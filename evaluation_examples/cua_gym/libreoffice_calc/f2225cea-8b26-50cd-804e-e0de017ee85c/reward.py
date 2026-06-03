"""
Reward Script: Fill blank cells in Customer Segment column by carrying values down,
               then create a pivot summary showing total orders per customer segment.
Task ID: osworld_calc_fill_blanks_above_005
Domain: libreoffice_calc
Scoring:
  Component 1: All blanks in column B (Customer Segment) are filled (0.5 pts)
  Component 2: Correct segment values carried down — spot check (0.2 pts)
  Component 3: 'Pivot Summary' sheet exists with correct headers (0.1 pts)
  Component 4: Pivot Summary has correct Total Amount per segment (0.2 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_blanks_above_005'

# Expected segment fill values after carrying down (row 2 onwards, 0-based index into data rows)
# Data rows 2..24 (24 rows total including header row 1)
EXPECTED_FILLED = {
    # Corporate: rows 2-7 (ORD-10021 to ORD-10026)
    2: 'Corporate', 3: 'Corporate', 4: 'Corporate', 5: 'Corporate', 6: 'Corporate', 7: 'Corporate',
    # Consumer: rows 8-13 (ORD-10027 to ORD-10032)
    8: 'Consumer', 9: 'Consumer', 10: 'Consumer', 11: 'Consumer', 12: 'Consumer', 13: 'Consumer',
    # Home Office: rows 14-18 (ORD-10033 to ORD-10037)
    14: 'Home Office', 15: 'Home Office', 16: 'Home Office', 17: 'Home Office', 18: 'Home Office',
    # Small Business: rows 19-24 (ORD-10038 to ORD-10043)
    19: 'Small Business', 20: 'Small Business', 21: 'Small Business',
    22: 'Small Business', 23: 'Small Business', 24: 'Small Business',
}

# Expected pivot summary: segment -> total amount (rounded to 2 decimal places)
EXPECTED_PIVOT = {
    'Consumer': 381.23,
    'Corporate': 4203.44,
    'Home Office': 1121.49,
    'Small Business': 747.25,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Orders' sheet must exist
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: 'Orders' sheet not found in workbook.")
        print("REWARD: 0.0")
        return 0.0

    ws_orders = wb['Orders']

    # Component 1: All blanks in column B (Customer Segment) are filled (0.5 pts)
    # This checks that every data row (rows 2-24) has a non-None, non-empty value in column B
    try:
        blank_count = 0
        total_data_rows = 0
        for row_idx in range(2, ws_orders.max_row + 1):
            cell_val = ws_orders.cell(row=row_idx, column=2).value
            total_data_rows += 1
            if cell_val is None or str(cell_val).strip() == '':
                blank_count += 1

        if blank_count == 0 and total_data_rows > 0:
            print(f"PASS: Component 1 — All {total_data_rows} data rows in column B are filled (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — {blank_count}/{total_data_rows} rows still have blanks in column B")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check column B blanks: {e}")

    # Component 2: Correct segment values carried down — spot check (0.2 pts)
    # Verify the correct segment was propagated to each row
    try:
        correct_fill = 0
        incorrect_rows = []
        for row_idx, expected_segment in EXPECTED_FILLED.items():
            actual_val = ws_orders.cell(row=row_idx, column=2).value
            if actual_val is not None and str(actual_val).strip() == expected_segment:
                correct_fill += 1
            else:
                incorrect_rows.append(f"row {row_idx}: expected '{expected_segment}', got '{actual_val}'")

        total_expected = len(EXPECTED_FILLED)
        if correct_fill == total_expected:
            print(f"PASS: Component 2 — All {total_expected} segment values correctly propagated (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Only {correct_fill}/{total_expected} rows have correct segment values")
            if incorrect_rows:
                for err in incorrect_rows[:5]:
                    print(f"  {err}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check segment fill values: {e}")

    # Component 3: 'Pivot Summary' sheet exists with correct headers (0.1 pts)
    try:
        pivot_sheet = None
        # Check both "Pivot Summary" and common variants
        for candidate in ['Pivot Summary', 'Sheet2', 'Summary', 'pivot_summary']:
            if candidate in wb.sheetnames:
                pivot_sheet = wb[candidate]
                break

        if pivot_sheet is not None:
            # Check headers in row 1
            h1 = pivot_sheet.cell(row=1, column=1).value
            h2 = pivot_sheet.cell(row=1, column=2).value
            h1_ok = h1 is not None and 'segment' in str(h1).lower()
            h2_ok = h2 is not None and ('amount' in str(h2).lower() or 'total' in str(h2).lower() or 'order' in str(h2).lower())

            if h1_ok and h2_ok:
                print(f"PASS: Component 3 — Pivot sheet '{pivot_sheet.title}' exists with valid headers: '{h1}', '{h2}' (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Pivot sheet found but headers wrong: col1='{h1}', col2='{h2}'")
        else:
            print(f"FAIL: Component 3 — No pivot/summary sheet found. Sheets present: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check pivot sheet: {e}")

    # Component 4: Pivot Summary has correct Total Amount per segment (0.2 pts)
    try:
        if pivot_sheet is None:
            print("FAIL: Component 4 — Skipped because no pivot sheet found")
        else:
            # Read segment->total from the pivot sheet
            pivot_data = {}
            for row_idx in range(2, pivot_sheet.max_row + 1):
                seg_val = pivot_sheet.cell(row=row_idx, column=1).value
                amt_val = pivot_sheet.cell(row=row_idx, column=2).value
                if seg_val is not None and amt_val is not None:
                    try:
                        pivot_data[str(seg_val).strip()] = round(float(amt_val), 2)
                    except (ValueError, TypeError):
                        pass

            # Check each expected segment total
            correct_totals = 0
            wrong_totals = []
            for seg, expected_total in EXPECTED_PIVOT.items():
                actual_total = pivot_data.get(seg)
                if actual_total is not None and abs(actual_total - expected_total) <= 0.02:
                    correct_totals += 1
                else:
                    wrong_totals.append(f"'{seg}': expected {expected_total}, got {actual_total}")

            if correct_totals == len(EXPECTED_PIVOT):
                print(f"PASS: Component 4 — All {len(EXPECTED_PIVOT)} segment totals correct in pivot sheet (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — Only {correct_totals}/{len(EXPECTED_PIVOT)} segment totals correct")
                for err in wrong_totals:
                    print(f"  {err}")
                print(f"  Pivot data found: {pivot_data}")
    except Exception as e:
        print(f"ERROR: Component 4 — Could not verify pivot totals: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
