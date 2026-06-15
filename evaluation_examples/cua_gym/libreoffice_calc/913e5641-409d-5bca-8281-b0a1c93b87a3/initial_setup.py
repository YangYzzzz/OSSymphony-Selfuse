"""
Initial Setup: Apply conditional formatting to Sales Amount column
Task ID: calc_ggf_001
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # Headers
    headers = ["Date", "Region", "Salesperson", "Amount"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data for 80 rows
    regions = ["Northeast", "Southeast", "Midwest", "West", "Southwest", "Pacific Northwest"]
    salespersons = [
        "Sarah Chen", "Marcus Johnson", "Emily Rivera", "David Kim",
        "Rachel Thompson", "James Williams", "Olivia Martinez", "Michael Brown",
        "Sophia Anderson", "Daniel Taylor", "Aisha Patel", "Robert Garcia",
        "Hannah Lee", "Christopher Wilson", "Isabella Moore", "Nathan Clark",
        "Victoria Adams", "Kevin Wright", "Lauren Scott", "Brian Hall"
    ]

    random.seed(42)  # reproducible

    # Generate 80 rows of data (rows 2 to 81)
    base_year = 2025
    for i in range(80):
        row = i + 2

        # Date: spread across Jan-Dec 2025
        month = (i % 12) + 1
        day = (i % 28) + 1
        date_str = f"{base_year}-{month:02d}-{day:02d}"
        ws.cell(row=row, column=1, value=date_str)

        # Region
        ws.cell(row=row, column=2, value=random.choice(regions))

        # Salesperson
        ws.cell(row=row, column=3, value=random.choice(salespersons))

        # Amount: range $800 to $12,400
        # Ensure a good mix of values above and below 5000
        amount = random.randint(800, 12400)
        ws.cell(row=row, column=4, value=amount)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 12

    # Number format for Amount column
    for row in range(2, 82):
        ws.cell(row=row, column=4).number_format = '$#,##0'

    # NO conditional formatting applied (that is the task)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
