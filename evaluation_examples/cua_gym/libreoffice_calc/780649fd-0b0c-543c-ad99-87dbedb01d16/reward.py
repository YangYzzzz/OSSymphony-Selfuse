"""
Reward Script: HR Headcount by Department Summary with COUNTIFS and Bar Chart
Task ID: calc_hr_headcount_by_dept_006
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.30): Summary sheet has correct headers (A1='Department', B1='Active Headcount')
                       and all 6 department names in A2:A7
  Component 2 (0.40): COUNTIFS formulas in B2:B7 referencing Employees sheet
                       with correct criteria for department and 'Active' status
  Component 3 (0.30): Bar/Column chart exists on Summary sheet with correct title
                       'Active Headcount by Department' and correct chart type

Total: 1.0
"""

import os
import openpyxl
from openpyxl.chart import BarChart

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_headcount_by_dept_006'

EXPECTED_DEPARTMENTS = ['Engineering', 'Marketing', 'Sales', 'HR', 'Finance', 'Operations']
EXPECTED_CHART_TITLE = 'Active Headcount by Department'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — fail fast if file is unreadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: Sheet 'Summary' not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Summary']

    # -----------------------------------------------------------------
    # Component 1: Headers and department names in Summary sheet (0.30)
    # Checks that A1='Department', B1='Active Headcount', and A2:A7 contain
    # all 6 expected department names. These are all absent in initial file.
    # -----------------------------------------------------------------
    try:
        a1_val = ws['A1'].value
        b1_val = ws['B1'].value

        header_ok = (
            a1_val is not None and str(a1_val).strip() == 'Department' and
            b1_val is not None and str(b1_val).strip() == 'Active Headcount'
        )

        dept_names_found = []
        for row in range(2, 8):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val is not None:
                dept_names_found.append(str(cell_val).strip())

        # Check all 6 departments are present (order may vary)
        all_depts_present = sorted(dept_names_found) == sorted(EXPECTED_DEPARTMENTS)

        if header_ok and all_depts_present:
            print(f"PASS: Component 1 — Headers correct (A1='Department', B1='Active Headcount')"
                  f" and all 6 departments found in A2:A7 (0.30 pts)")
            total_score += 0.30
        elif header_ok and not all_depts_present:
            print(f"FAIL: Component 1 — Headers correct, but department names mismatch."
                  f" Found: {dept_names_found}, Expected: {EXPECTED_DEPARTMENTS}")
        elif not header_ok and all_depts_present:
            print(f"FAIL: Component 1 — Department names OK, but headers wrong."
                  f" A1={repr(a1_val)}, B1={repr(b1_val)}")
        else:
            print(f"FAIL: Component 1 — Headers wrong (A1={repr(a1_val)}, B1={repr(b1_val)})"
                  f" and departments not found.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------
    # Component 2: COUNTIFS formulas in B2:B7 referencing Employees sheet (0.40)
    # Each cell B2:B7 must contain a COUNTIFS formula referencing:
    #   - Employees column C (department) matched against cell in column A same row
    #   - Employees column D (status) matched against "Active"
    # -----------------------------------------------------------------
    try:
        formula_count = 0
        countifs_errors = []

        for row in range(2, 8):
            cell = ws.cell(row=row, column=2)
            val = cell.value

            if val is None:
                countifs_errors.append(f"B{row}: empty")
                continue

            val_str = str(val).strip().upper().replace(' ', '')

            # Must be a formula starting with =COUNTIFS
            if not val_str.startswith('=COUNTIFS'):
                countifs_errors.append(f"B{row}: not a COUNTIFS formula: {repr(val)}")
                continue

            # Must reference Employees column C (department)
            employees_c_ref = 'EMPLOYEES' in val_str.upper() and (
                'C:C' in val_str.upper() or 'EMPLOYEES.C' in val_str.upper()
            )

            # Must reference Employees column D (status) with "Active"
            employees_d_ref = 'EMPLOYEES' in val_str.upper() and (
                'D:D' in val_str.upper() or 'EMPLOYEES.D' in val_str.upper()
            )

            has_active_criterion = '"ACTIVE"' in val_str.upper()

            if employees_c_ref and employees_d_ref and has_active_criterion:
                formula_count += 1
            else:
                countifs_errors.append(
                    f"B{row}: COUNTIFS formula missing required references. "
                    f"emp_c={employees_c_ref}, emp_d={employees_d_ref}, active={has_active_criterion}. "
                    f"Formula: {repr(val)}"
                )

        if formula_count == 6:
            print(f"PASS: Component 2 — All 6 COUNTIFS formulas in B2:B7 correctly reference"
                  f" Employees columns C and D with 'Active' criterion (0.40 pts)")
            total_score += 0.40
        elif formula_count >= 3:
            partial = round(0.40 * formula_count / 6, 2)
            print(f"PARTIAL: Component 2 — {formula_count}/6 COUNTIFS formulas correct."
                  f" Errors: {countifs_errors}")
            # Only full credit for this component — no partial within
            print(f"  (Partial credit NOT awarded — require all 6 correct)")
        else:
            print(f"FAIL: Component 2 — Only {formula_count}/6 COUNTIFS formulas correct."
                  f" Errors: {countifs_errors}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------
    # Component 3: Bar/Column chart on Summary sheet with correct title (0.30)
    # Chart must:
    #   - Exist on Summary sheet (at least 1 chart)
    #   - Be a BarChart type (bar or col sub-type)
    #   - Have title 'Active Headcount by Department'
    # -----------------------------------------------------------------
    try:
        charts = ws._charts

        if not charts:
            print("FAIL: Component 3 — No charts found on Summary sheet")
        else:
            chart = charts[0]

            # Check chart is BarChart (bar or column)
            is_bar_or_col = isinstance(chart, BarChart)

            # Extract chart title text
            chart_title_text = None
            try:
                title_obj = chart.title
                if title_obj is not None:
                    rich = title_obj.tx.rich
                    for p in rich.p:
                        for r in p.r:
                            chart_title_text = r.t
                            break
                        if chart_title_text:
                            break
            except Exception:
                chart_title_text = None

            title_correct = (
                chart_title_text is not None and
                chart_title_text.strip() == EXPECTED_CHART_TITLE
            )

            if is_bar_or_col and title_correct:
                print(f"PASS: Component 3 — Bar/Column chart found on Summary sheet with"
                      f" correct title '{chart_title_text}' (0.30 pts)")
                total_score += 0.30
            elif is_bar_or_col and not title_correct:
                print(f"FAIL: Component 3 — Chart is BarChart but title incorrect."
                      f" Found: {repr(chart_title_text)}, Expected: '{EXPECTED_CHART_TITLE}'")
            elif not is_bar_or_col and title_correct:
                print(f"FAIL: Component 3 — Title correct but chart type is not BarChart."
                      f" Found: {type(chart).__name__}")
            else:
                print(f"FAIL: Component 3 — Chart type wrong ({type(chart).__name__})"
                      f" and title wrong ({repr(chart_title_text)})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
