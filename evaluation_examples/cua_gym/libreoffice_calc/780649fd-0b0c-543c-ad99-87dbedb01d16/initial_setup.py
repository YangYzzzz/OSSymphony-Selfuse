"""
Initial Setup: HR Headcount by Department
Task ID: calc_hr_headcount_by_dept_006
Domain: libreoffice_calc

Creates an HR employee spreadsheet with:
- Sheet 'Employees': 142 rows of employee data (ID, Name, Department, Status)
- Sheet 'Summary': empty sheet (task requires adding COUNTIFS + chart)
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_headcount_by_dept_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Employees ---
    ws_emp = wb.active
    ws_emp.title = 'Employees'

    # Headers
    headers = ['Emp ID', 'Name', 'Department', 'Status']
    for col, h in enumerate(headers, 1):
        ws_emp.cell(row=1, column=col, value=h)

    # Realistic employee data: 142 employees
    # Departments: Engineering, Marketing, Sales, HR, Finance, Operations
    # Statuses: Active, Inactive, On Leave
    employees = [
        # Engineering (35 employees)
        ('E001', 'Sarah Chen', 'Engineering', 'Active'),
        ('E002', 'Marcus Johnson', 'Engineering', 'Active'),
        ('E003', 'Priya Patel', 'Engineering', 'Active'),
        ('E004', 'David Kim', 'Engineering', 'Active'),
        ('E005', 'Rachel Torres', 'Engineering', 'Active'),
        ('E006', 'James Williams', 'Engineering', 'Inactive'),
        ('E007', 'Emily Zhang', 'Engineering', 'Active'),
        ('E008', 'Carlos Mendez', 'Engineering', 'Active'),
        ('E009', 'Jessica Park', 'Engineering', 'On Leave'),
        ('E010', 'Ryan Mitchell', 'Engineering', 'Active'),
        ('E011', 'Aisha Okafor', 'Engineering', 'Active'),
        ('E012', 'Nathan Brooks', 'Engineering', 'Active'),
        ('E013', 'Mei-Ling Wu', 'Engineering', 'Active'),
        ('E014', 'Tyler Anderson', 'Engineering', 'Inactive'),
        ('E015', 'Sofia Alvarez', 'Engineering', 'Active'),
        ('E016', 'Kevin O\'Brien', 'Engineering', 'Active'),
        ('E017', 'Fatima Hassan', 'Engineering', 'Active'),
        ('E018', 'Derek Thompson', 'Engineering', 'Active'),
        ('E019', 'Yuki Tanaka', 'Engineering', 'On Leave'),
        ('E020', 'Brittany Lewis', 'Engineering', 'Active'),
        ('E021', 'Mohammed Al-Rashid', 'Engineering', 'Active'),
        ('E022', 'Chloe Martin', 'Engineering', 'Active'),
        ('E023', 'Andre Robinson', 'Engineering', 'Active'),
        ('E024', 'Ingrid Johansson', 'Engineering', 'Active'),
        ('E025', 'Jamal Washington', 'Engineering', 'Active'),
        ('E026', 'Laura Garcia', 'Engineering', 'Inactive'),
        ('E027', 'Steve Nguyen', 'Engineering', 'Active'),
        ('E028', 'Nina Petrov', 'Engineering', 'Active'),
        ('E029', 'Omar Farouq', 'Engineering', 'Active'),
        ('E030', 'Samantha Price', 'Engineering', 'Active'),
        ('E031', 'Ben Kowalski', 'Engineering', 'Active'),
        ('E032', 'Amara Diallo', 'Engineering', 'Active'),
        ('E033', 'Ethan Clark', 'Engineering', 'On Leave'),
        ('E034', 'Hana Nakamura', 'Engineering', 'Active'),
        ('E035', 'Leo Fernandez', 'Engineering', 'Active'),
        # Marketing (22 employees)
        ('M001', 'Amanda Foster', 'Marketing', 'Active'),
        ('M002', 'Brandon Scott', 'Marketing', 'Active'),
        ('M003', 'Christina Lee', 'Marketing', 'Active'),
        ('M004', 'Daniel Brown', 'Marketing', 'Inactive'),
        ('M005', 'Elena Volkov', 'Marketing', 'Active'),
        ('M006', 'Frank Davis', 'Marketing', 'Active'),
        ('M007', 'Grace Kim', 'Marketing', 'Active'),
        ('M008', 'Henry Wilson', 'Marketing', 'On Leave'),
        ('M009', 'Isabella Martinez', 'Marketing', 'Active'),
        ('M010', 'Jake Taylor', 'Marketing', 'Active'),
        ('M011', 'Kara White', 'Marketing', 'Active'),
        ('M012', 'Liam Harris', 'Marketing', 'Active'),
        ('M013', 'Maya Singh', 'Marketing', 'Active'),
        ('M014', 'Noah Jackson', 'Marketing', 'Inactive'),
        ('M015', 'Olivia Thomas', 'Marketing', 'Active'),
        ('M016', 'Patrick Moore', 'Marketing', 'Active'),
        ('M017', 'Quinn Robinson', 'Marketing', 'Active'),
        ('M018', 'Rebecca Hall', 'Marketing', 'Active'),
        ('M019', 'Samuel Young', 'Marketing', 'On Leave'),
        ('M020', 'Tara Allen', 'Marketing', 'Active'),
        ('M021', 'Ulysses King', 'Marketing', 'Active'),
        ('M022', 'Vera Wright', 'Marketing', 'Active'),
        # Sales (28 employees)
        ('S001', 'Aaron Scott', 'Sales', 'Active'),
        ('S002', 'Beverly Green', 'Sales', 'Active'),
        ('S003', 'Charles Adams', 'Sales', 'Active'),
        ('S004', 'Diana Carter', 'Sales', 'Inactive'),
        ('S005', 'Edward Mitchell', 'Sales', 'Active'),
        ('S006', 'Fiona Nelson', 'Sales', 'Active'),
        ('S007', 'George Bell', 'Sales', 'Active'),
        ('S008', 'Hannah Cooper', 'Sales', 'Active'),
        ('S009', 'Ivan Reed', 'Sales', 'On Leave'),
        ('S010', 'Julia Bailey', 'Sales', 'Active'),
        ('S011', 'Karl Rivera', 'Sales', 'Active'),
        ('S012', 'Linda Cox', 'Sales', 'Active'),
        ('S013', 'Michael Ward', 'Sales', 'Active'),
        ('S014', 'Nancy Torres', 'Sales', 'Active'),
        ('S015', 'Oscar Powell', 'Sales', 'Inactive'),
        ('S016', 'Paula Russell', 'Sales', 'Active'),
        ('S017', 'Quincy Barnes', 'Sales', 'Active'),
        ('S018', 'Rose Henderson', 'Sales', 'Active'),
        ('S019', 'Simon Coleman', 'Sales', 'Active'),
        ('S020', 'Tiffany Jenkins', 'Sales', 'On Leave'),
        ('S021', 'Umar Perry', 'Sales', 'Active'),
        ('S022', 'Valerie Patterson', 'Sales', 'Active'),
        ('S023', 'Walter Hughes', 'Sales', 'Active'),
        ('S024', 'Xena Flores', 'Sales', 'Active'),
        ('S025', 'Yolanda Griffin', 'Sales', 'Active'),
        ('S026', 'Zachary Diaz', 'Sales', 'Inactive'),
        ('S027', 'Alexa Myers', 'Sales', 'Active'),
        ('S028', 'Bryan Ford', 'Sales', 'Active'),
        # HR (14 employees)
        ('H001', 'Catherine Hamilton', 'HR', 'Active'),
        ('H002', 'Douglas Graham', 'HR', 'Active'),
        ('H003', 'Eleanor Sullivan', 'HR', 'Active'),
        ('H004', 'Felix Wallace', 'HR', 'Inactive'),
        ('H005', 'Gloria Woods', 'HR', 'Active'),
        ('H006', 'Herbert Cole', 'HR', 'Active'),
        ('H007', 'Irene West', 'HR', 'On Leave'),
        ('H008', 'Jerome Jordan', 'HR', 'Active'),
        ('H009', 'Kathleen Berry', 'HR', 'Active'),
        ('H010', 'Lawrence Gibson', 'HR', 'Active'),
        ('H011', 'Miriam Hunt', 'HR', 'Active'),
        ('H012', 'Nicholas Spencer', 'HR', 'Active'),
        ('H013', 'Ophelia Stone', 'HR', 'Inactive'),
        ('H014', 'Philip Burke', 'HR', 'Active'),
        # Finance (18 employees)
        ('F001', 'Abigail Reid', 'Finance', 'Active'),
        ('F002', 'Benjamin Fox', 'Finance', 'Active'),
        ('F003', 'Cecilia Murray', 'Finance', 'Active'),
        ('F004', 'Dennis Fischer', 'Finance', 'Inactive'),
        ('F005', 'Elise Crawford', 'Finance', 'Active'),
        ('F006', 'Francis Lane', 'Finance', 'Active'),
        ('F007', 'Gabrielle Lynch', 'Finance', 'Active'),
        ('F008', 'Harold Palmer', 'Finance', 'On Leave'),
        ('F009', 'Irma Carr', 'Finance', 'Active'),
        ('F010', 'Jonathan Wolfe', 'Finance', 'Active'),
        ('F011', 'Kimberly Horton', 'Finance', 'Active'),
        ('F012', 'Leonard Ortega', 'Finance', 'Active'),
        ('F013', 'Margaret Porter', 'Finance', 'Active'),
        ('F014', 'Norman Freeman', 'Finance', 'Inactive'),
        ('F015', 'Penelope Dean', 'Finance', 'Active'),
        ('F016', 'Randolph Hicks', 'Finance', 'Active'),
        ('F017', 'Sandra Weaver', 'Finance', 'Active'),
        ('F018', 'Theodore Bowman', 'Finance', 'Active'),
        # Operations (25 employees)
        ('O001', 'Alice Barker', 'Operations', 'Active'),
        ('O002', 'Bruce Hudson', 'Operations', 'Active'),
        ('O003', 'Claire Miles', 'Operations', 'Active'),
        ('O004', 'Donald Norris', 'Operations', 'Inactive'),
        ('O005', 'Edna Hanson', 'Operations', 'Active'),
        ('O006', 'Floyd Page', 'Operations', 'Active'),
        ('O007', 'Gail Warren', 'Operations', 'Active'),
        ('O008', 'Horace Cunningham', 'Operations', 'On Leave'),
        ('O009', 'Inez Dunn', 'Operations', 'Active'),
        ('O010', 'Jeffrey Garrett', 'Operations', 'Active'),
        ('O011', 'Kirsten Larson', 'Operations', 'Active'),
        ('O012', 'Lester Brewer', 'Operations', 'Active'),
        ('O013', 'Marilyn Nichols', 'Operations', 'Active'),
        ('O014', 'Neil Pearce', 'Operations', 'Inactive'),
        ('O015', 'Ora Stanley', 'Operations', 'Active'),
        ('O016', 'Percy Holland', 'Operations', 'Active'),
        ('O017', 'Quentin Chandler', 'Operations', 'Active'),
        ('O018', 'Roberta Sutton', 'Operations', 'Active'),
        ('O019', 'Scott Curry', 'Operations', 'On Leave'),
        ('O020', 'Tammy Holt', 'Operations', 'Active'),
        ('O021', 'Ulric Santos', 'Operations', 'Active'),
        ('O022', 'Velma Lowe', 'Operations', 'Active'),
        ('O023', 'Wayne Barber', 'Operations', 'Active'),
        ('O024', 'Xenia Hardy', 'Operations', 'Active'),
        ('O025', 'Yves Caldwell', 'Operations', 'Inactive'),
    ]

    # Write employee data starting from row 2
    for r, (emp_id, name, dept, status) in enumerate(employees, 2):
        ws_emp.cell(row=r, column=1, value=emp_id)
        ws_emp.cell(row=r, column=2, value=name)
        ws_emp.cell(row=r, column=3, value=dept)
        ws_emp.cell(row=r, column=4, value=status)

    # Set column widths for readability
    ws_emp.column_dimensions['A'].width = 10
    ws_emp.column_dimensions['B'].width = 25
    ws_emp.column_dimensions['C'].width = 15
    ws_emp.column_dimensions['D'].width = 12

    # --- Sheet 2: Summary (empty — task will add COUNTIFS + chart) ---
    ws_summary = wb.create_sheet('Summary')
    # Leave completely empty — task requires populating this sheet

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Employees sheet: {len(employees)} rows (rows 2-{len(employees)+1})')
    print('Summary sheet: empty (ready for task)')


create_initial()
