"""
Reward Script: Minimize total shipping cost using Solver on transportation problem
Task ID: calc_gg5_027
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Non-zero shipment quantities exist in B2:E5
  Component 2 (0.10): All shipment quantities >= 0
  Component 3 (0.25): Supply constraints satisfied (row sums <= limits in G2:G5)
  Component 4 (0.25): Demand constraints satisfied (col sums >= requirements in B7:E7)
  Component 5 (0.20): Total cost is optimal (near 10050)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_027'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Routes' not in wb.sheetnames:
        print("CRITICAL: 'Routes' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Routes']

    # Read shipment quantities from B2:E5
    shipments = []
    try:
        for r in range(2, 6):
            row_vals = []
            for c in range(2, 6):  # B=2, C=3, D=4, E=5
                val = ws.cell(row=r, column=c).value
                if val is None:
                    val = 0
                row_vals.append(float(val))
            shipments.append(row_vals)
        print(f"INFO: Shipment matrix: {shipments}")
    except Exception as e:
        print(f"CRITICAL: Cannot read shipment quantities B2:E5: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Read supply limits from G2:G5
    supply_limits = []
    try:
        for r in range(2, 6):
            val = ws.cell(row=r, column=7).value  # G column
            supply_limits.append(float(val))
        print(f"INFO: Supply limits: {supply_limits}")
    except Exception as e:
        print(f"CRITICAL: Cannot read supply limits G2:G5: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Read demand requirements from B7:E7
    demand_reqs = []
    try:
        for c in range(2, 6):
            val = ws.cell(row=7, column=c).value
            demand_reqs.append(float(val))
        print(f"INFO: Demand requirements: {demand_reqs}")
    except Exception as e:
        print(f"CRITICAL: Cannot read demand requirements B7:E7: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Read cost matrix from J2:M5
    costs = []
    try:
        for r in range(2, 6):
            row_vals = []
            for c in range(10, 14):  # J=10, K=11, L=12, M=13
                val = ws.cell(row=r, column=c).value
                row_vals.append(float(val))
            costs.append(row_vals)
        print(f"INFO: Cost matrix: {costs}")
    except Exception as e:
        print(f"CRITICAL: Cannot read cost matrix J2:M5: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Non-zero shipment quantities exist in B2:E5 (0.20 points)
    # In initial state all values are 0. After solver runs, at least some must be > 0.
    try:
        total_shipped = sum(sum(row) for row in shipments)
        if total_shipped > 0:
            print(f"PASS: Component 1 — Non-zero shipments found, total shipped={total_shipped} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — All shipment quantities are 0, solver has not been run")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All shipment quantities >= 0 (non-negativity constraint) (0.10 points)
    # Only award if Component 1 passed (otherwise trivially true for all-zero initial state)
    try:
        if total_shipped > 0:
            all_non_neg = all(shipments[i][j] >= 0 for i in range(4) for j in range(4))
            if all_non_neg:
                print(f"PASS: Component 2 — All quantities >= 0 (0.10 pts)")
                total_score += 0.10
            else:
                neg_cells = [(i, j, shipments[i][j]) for i in range(4) for j in range(4) if shipments[i][j] < 0]
                print(f"FAIL: Component 2 — Negative quantities found: {neg_cells}")
        else:
            print(f"FAIL: Component 2 — Skipped (no shipments placed)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Supply constraints satisfied (0.25 points)
    # Each row sum of shipments must be <= corresponding supply limit
    try:
        if total_shipped > 0:
            supply_ok_count = 0
            for i in range(4):
                row_sum = sum(shipments[i])
                if row_sum <= supply_limits[i] + 0.01:  # small tolerance
                    supply_ok_count += 1
                else:
                    print(f"FAIL: Component 3 — Warehouse {i} ships {row_sum} > limit {supply_limits[i]}")
            if supply_ok_count == 4:
                print(f"PASS: Component 3 — All 4 supply constraints satisfied (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Only {supply_ok_count}/4 supply constraints satisfied")
        else:
            print(f"FAIL: Component 3 — Skipped (no shipments placed)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Demand constraints satisfied (0.25 points)
    # Each column sum of shipments must be >= corresponding demand requirement
    try:
        if total_shipped > 0:
            demand_ok_count = 0
            for j in range(4):
                col_sum = sum(shipments[i][j] for i in range(4))
                if col_sum >= demand_reqs[j] - 0.01:  # small tolerance
                    demand_ok_count += 1
                else:
                    print(f"FAIL: Component 4 — DC {j} receives {col_sum} < required {demand_reqs[j]}")
            if demand_ok_count == 4:
                print(f"PASS: Component 4 — All 4 demand constraints satisfied (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 4 — Only {demand_ok_count}/4 demand constraints satisfied")
        else:
            print(f"FAIL: Component 4 — Skipped (no shipments placed)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Total cost is optimal or near-optimal (0.20 points)
    # The optimal total cost for this problem is 10050.
    # We compute SUMPRODUCT(costs, shipments) since openpyxl can't evaluate formulas.
    try:
        if total_shipped > 0:
            computed_cost = sum(
                shipments[i][j] * costs[i][j]
                for i in range(4) for j in range(4)
            )
            optimal_cost = 10050.0
            print(f"INFO: Computed total cost = {computed_cost}, optimal = {optimal_cost}")

            # Accept cost within 5% of optimal (solver may find slightly different feasible solutions)
            if abs(computed_cost - optimal_cost) <= optimal_cost * 0.05:
                print(f"PASS: Component 5 — Cost {computed_cost} is near-optimal (within 5% of {optimal_cost}) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 5 — Cost {computed_cost} is not near optimal {optimal_cost}")
        else:
            print(f"FAIL: Component 5 — Skipped (no shipments placed)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
