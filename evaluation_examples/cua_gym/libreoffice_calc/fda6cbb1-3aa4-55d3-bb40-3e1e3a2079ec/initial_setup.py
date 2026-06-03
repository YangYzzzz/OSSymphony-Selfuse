"""
Initial Setup: Transportation problem for Solver optimization
Task ID: calc_gg5_027
Domain: libreoffice_calc

Creates logistics.xlsx with a Routes sheet containing a transportation problem:
- 4 supply nodes (warehouses) x 4 demand nodes (distribution centers)
- Unit cost matrix in J1:M5
- Decision variables (shipment quantities) in B2:E5, all set to 0
- Objective formula in H1: =SUMPRODUCT of costs and quantities
- Supply remaining formulas in F2:F5
- Demand unmet formulas in B6:E6
- Constraints are implicit in the formulas for Solver setup
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ===== Sheet: Routes =====
    ws = wb.active
    ws.title = 'Routes'

    # --- Styles ---
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    data_font = Font(name='Calibri', size=11)
    cost_header_fill = PatternFill(start_color='FF548235', end_color='FF548235', fill_type='solid')
    label_fill = PatternFill(start_color='FFD6E4F0', end_color='FFD6E4F0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Layout ---
    # Columns: A=Source labels, B-E=Dest cities, F=Supply Remaining, G=blank, H=Objective
    # Row 1: Headers
    # Rows 2-5: Warehouses (supply nodes)
    # Row 6: Demand row

    # Supply data
    supply_names = ['Portland Warehouse', 'Denver Warehouse', 'Atlanta Warehouse', 'Chicago Warehouse']
    supply_limits = [250, 350, 400, 200]

    # Demand data
    demand_names = ['New York DC', 'Los Angeles DC', 'Dallas DC', 'Miami DC']
    demand_values = [300, 250, 350, 300]

    # Unit cost matrix (rows=supply, cols=demand)
    cost_matrix = [
        [8, 6, 10, 9],
        [9, 12, 13, 7],
        [14, 9, 16, 5],
        [18, 20, 12, 10],
    ]

    # --- Row 1: Column headers for shipment table ---
    ws.cell(row=1, column=1, value='Source / Destination').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
    ws.cell(row=1, column=1).border = thin_border

    for j, dest in enumerate(demand_names, 2):  # B1:E1
        cell = ws.cell(row=1, column=j, value=dest)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # F1: Supply Remaining header
    cell = ws.cell(row=1, column=6, value='Supply Remaining')
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

    # G1: Supply Limit header
    cell = ws.cell(row=1, column=7, value='Supply Limit')
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

    # H1: Objective formula (SUMPRODUCT of cost matrix and shipment quantities)
    # Cost matrix is in J2:M5, quantities in B2:E5
    ws.cell(row=1, column=8, value='=SUMPRODUCT(J2:M5,B2:E5)')
    ws.cell(row=1, column=8).font = Font(name='Calibri', size=14, bold=True, color='FF0000')
    ws.cell(row=1, column=8).number_format = '$#,##0'
    ws.cell(row=1, column=8).border = thin_border

    # H2: Label for objective
    ws.cell(row=2, column=8, value='Total Shipping Cost (Minimize)')
    ws.cell(row=2, column=8).font = Font(name='Calibri', size=9, italic=True)

    # --- Rows 2-5: Supply nodes with shipment quantities = 0 ---
    for i in range(4):
        row = i + 2

        # A: Source name
        cell = ws.cell(row=row, column=1, value=supply_names[i])
        cell.font = data_font
        cell.fill = label_fill
        cell.border = thin_border

        # B-E: Shipment quantities (decision variables), all 0
        for j in range(4):
            col = j + 2
            cell = ws.cell(row=row, column=col, value=0)
            cell.font = data_font
            cell.number_format = '0'
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

        # F: Supply remaining = Supply limit - SUM(shipments in this row)
        # e.g., F2 = G2 - SUM(B2:E2)
        col_letter_b = 'B'
        col_letter_e = 'E'
        cell = ws.cell(row=row, column=6,
                       value=f'=G{row}-SUM({col_letter_b}{row}:{col_letter_e}{row})')
        cell.font = data_font
        cell.number_format = '0'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

        # G: Supply limit (constant)
        cell = ws.cell(row=row, column=7, value=supply_limits[i])
        cell.font = data_font
        cell.number_format = '0'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # --- Row 6: Demand row ---
    cell = ws.cell(row=6, column=1, value='Demand Unmet')
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

    for j in range(4):
        col = j + 2
        col_letter = chr(ord('B') + j)
        # Demand unmet = Demand requirement - SUM(shipments in this column)
        # Demand requirements are in row 7
        cell = ws.cell(row=6, column=col,
                       value=f'={col_letter}7-SUM({col_letter}2:{col_letter}5)')
        cell.font = data_font
        cell.number_format = '0'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Row 7: Demand requirements (constants)
    cell = ws.cell(row=7, column=1, value='Demand Requirement')
    cell.font = Font(name='Calibri', size=11, bold=True)
    cell.fill = label_fill
    cell.border = thin_border

    for j in range(4):
        col = j + 2
        cell = ws.cell(row=7, column=col, value=demand_values[j])
        cell.font = data_font
        cell.number_format = '0'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # --- Cost Matrix (J1:M5) ---
    # Header row
    cell = ws.cell(row=1, column=10, value='Unit Cost Matrix ($/unit)')
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.fill = cost_header_fill
    ws.merge_cells('J1:M1')
    cell.alignment = Alignment(horizontal='center')

    for i in range(4):
        row = i + 2
        for j in range(4):
            col = j + 10  # J=10, K=11, L=12, M=13
            cell = ws.cell(row=row, column=col, value=cost_matrix[i][j])
            cell.font = data_font
            cell.number_format = '$#,##0'
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    # Cost matrix row labels (I2:I5)
    for i, name in enumerate(supply_names, 2):
        cell = ws.cell(row=i, column=9, value=name)
        cell.font = Font(name='Calibri', size=9, italic=True)

    # Cost matrix column labels (J6:M6) — destination names
    for j, dest in enumerate(demand_names, 10):
        cell = ws.cell(row=6, column=j, value=dest)
        cell.font = Font(name='Calibri', size=9, italic=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 22
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 12
    ws.column_dimensions['M'].width = 12

    # ===== Sheet: Instructions =====
    ws2 = wb.create_sheet('Instructions')
    ws2['A1'] = 'Transportation Problem — Solver Setup Guide'
    ws2['A1'].font = Font(size=14, bold=True)
    ws2['A3'] = 'Objective Cell: H1 (Total Shipping Cost) — Minimize'
    ws2['A4'] = 'Variable Cells: B2:E5 (Shipment Quantities)'
    ws2['A5'] = 'Constraints:'
    ws2['A6'] = '  1. F2:F5 >= 0 (Supply remaining must be non-negative)'
    ws2['A7'] = '  2. B6:E6 >= 0 (Demand unmet must be non-negative — actually = 0 for feasibility)'
    ws2['A8'] = '  3. B2:E5 >= 0 (Non-negativity of shipment quantities)'
    ws2['A10'] = 'This is a balanced transportation problem: Total Supply = Total Demand = 1,200 units'
    ws2['A11'] = 'Use Tools > Solver to find the optimal shipping plan.'
    ws2.column_dimensions['A'].width = 80

    # --- Save ---
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # --- GUI-ready: open in LibreOffice Calc ---
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
