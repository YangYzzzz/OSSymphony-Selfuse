"""
Initial Setup: HR Salary Increase History
Task ID: calc_hr_salary_increase_history_071
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_increase_history_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Salary History'

    # Headers
    headers = ['Emp ID', 'Name', 'Original Salary', 'Current Salary', 'Last Raise Date', 'Total Increase %', 'Years Since Raise']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Realistic employee data: 95 rows (rows 2-96)
    first_names = [
        'James', 'Mary', 'Robert', 'Patricia', 'John', 'Jennifer', 'Michael', 'Linda',
        'William', 'Barbara', 'David', 'Elizabeth', 'Richard', 'Susan', 'Joseph', 'Jessica',
        'Thomas', 'Sarah', 'Charles', 'Karen', 'Christopher', 'Lisa', 'Daniel', 'Nancy',
        'Matthew', 'Betty', 'Anthony', 'Margaret', 'Mark', 'Sandra', 'Donald', 'Ashley',
        'Steven', 'Dorothy', 'Paul', 'Kimberly', 'Andrew', 'Emily', 'Joshua', 'Donna',
        'Kenneth', 'Michelle', 'Kevin', 'Carol', 'Brian', 'Amanda', 'George', 'Melissa',
        'Timothy', 'Deborah', 'Ronald', 'Stephanie', 'Edward', 'Rebecca', 'Jason', 'Sharon',
        'Jeffrey', 'Laura', 'Ryan', 'Cynthia', 'Jacob', 'Kathleen', 'Gary', 'Amy',
        'Nicholas', 'Angela', 'Eric', 'Shirley', 'Jonathan', 'Anna', 'Stephen', 'Brenda',
        'Larry', 'Pamela', 'Justin', 'Emma', 'Scott', 'Nicole', 'Brandon', 'Helen',
        'Benjamin', 'Samantha', 'Samuel', 'Katherine', 'Raymond', 'Christine', 'Gregory', 'Debra',
        'Frank', 'Rachel', 'Alexander', 'Carolyn'
    ]

    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
        'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson', 'Anderson', 'Thomas',
        'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'Perez', 'Thompson', 'White',
        'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young',
        'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores',
        'Green', 'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts', 'Chen', 'Patel', 'Kim', 'Murphy', 'Watson', 'Brooks',
        'Kelly', 'Howard', 'Ward', 'Cox', 'Diaz', 'Richardson', 'Wood', 'Watson',
        'Brooks', 'Bennett', 'Gray', 'James', 'Reyes', 'Cruz', 'Hughes', 'Price',
        'Myers', 'Long', 'Foster', 'Sanders', 'Ross', 'Morales', 'Powell', 'Sullivan',
        'Russell', 'Ortiz', 'Jenkins', 'Gutierrez', 'Perry', 'Butler', 'Barnes', 'Fisher',
        'Henderson', 'Coleman', 'Simmons', 'Patterson', 'Jordan', 'Reynolds'
    ]

    departments = ['Engineering', 'Marketing', 'Finance', 'HR', 'Sales', 'Operations', 'Legal', 'Product', 'Design', 'IT']

    random.seed(42)

    today = date(2026, 3, 4)
    hire_start = date(2015, 1, 1)
    hire_end = date(2024, 12, 31)

    for i in range(95):
        emp_id = f'EMP{1001 + i:04d}'
        name = f'{first_names[i % len(first_names)]} {last_names[i % len(last_names)]}'

        # Original salary: hired between 2015-2024
        hire_days = (hire_end - hire_start).days
        hire_date = hire_start + timedelta(days=random.randint(0, hire_days))

        # Original salary range 40000-90000
        original_salary = round(random.randint(40000, 90000) / 100) * 100

        # Current salary is original + some increases (5%-35% more)
        increase_pct = random.uniform(0.05, 0.35)
        current_salary = round(original_salary * (1 + increase_pct) / 100) * 100

        # Last raise date: between hire date and today, range 6 months to 5 years ago
        max_raise_ago = min((today - hire_date).days, 365 * 5)
        min_raise_ago = 180
        if max_raise_ago < min_raise_ago:
            max_raise_ago = min_raise_ago + 1
        raise_days_ago = random.randint(min_raise_ago, max_raise_ago)
        last_raise_date = today - timedelta(days=raise_days_ago)

        row = i + 2
        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=original_salary)
        ws.cell(row=row, column=4, value=current_salary)
        ws.cell(row=row, column=5, value=last_raise_date)
        ws.cell(row=row, column=5).number_format = 'yyyy-mm-dd'
        # Columns F (Total Increase %) and G (Years Since Raise) are intentionally EMPTY

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Salary History')
    print(f'  Headers in row 1: A-G')
    print(f'  Data rows: 2-96 (95 employees)')
    print(f'  Columns F and G are EMPTY (to be filled by the agent)')

create_initial()
