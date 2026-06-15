"""
Reward Script: HR Salary Increase History with Cumulative % and Years Since Raise
Task ID: calc_hr_salary_increase_history_071
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): F2:F96 contain formula =(D{row}-C{row})/C{row}
  - Component 2 (0.20): F2:F96 formatted as percentage 0.0%
  - Component 3 (0.30): G2:G96 contain formula =DATEDIF(E{row},TODAY(),"Y")
  - Component 4 (0.10): G2:G96 formatted as integer (0)
  - Component 5 (0.10): Conditional formatting on A2:G96 with amber (#FFC000) fill when G>=2
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_salary_increase_history_071'


def normalize_formula(formula):
    """Normalize formula string for comparison - strip spaces, uppercase."""
    if not isinstance(formula, str):
        return ''
    return formula.strip().upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Salary History' sheet must exist
    if 'Salary History' not in wb.sheetnames:
        print("CRITICAL: 'Salary History' sheet not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Salary History']

    # Component 1: F2:F96 contain formula =(D{row}-C{row})/C{row} (0.30 points)
    # This checks the cumulative salary increase percentage formula
    try:
        f_formula_correct = 0
        f_formula_total = 95  # rows 2 through 96
        f_missing = []
        for row in range(2, 97):
            cell_val = ws.cell(row=row, column=6).value  # Column F
            if cell_val is None:
                f_missing.append(row)
                continue
            formula_norm = normalize_formula(str(cell_val))
            expected_norm = normalize_formula(f'=(D{row}-C{row})/C{row}')
            if formula_norm == expected_norm:
                f_formula_correct += 1
            else:
                if len(f_missing) == 0 and f_formula_correct == 0:
                    # Print first mismatch for debugging
                    print(f"FAIL sample: F{row} has {repr(cell_val)}, expected =(D{row}-C{row})/C{row}")

        ratio = f_formula_correct / f_formula_total
        if ratio >= 0.95:
            print(f"PASS: Component 1 — F column formulas correct ({f_formula_correct}/{f_formula_total}) (0.30 pts)")
            total_score += 0.30
        elif ratio >= 0.5:
            partial = round(0.15 * ratio / 0.95, 2)
            print(f"PARTIAL: Component 1 — F column formulas partially correct ({f_formula_correct}/{f_formula_total}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — F column formulas largely missing/incorrect ({f_formula_correct}/{f_formula_total})")
            if f_missing:
                print(f"  Missing rows (up to 5): {f_missing[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2:F96 formatted as percentage 0.0% (0.20 points)
    # Check that cells in column F have the 0.0% number format
    try:
        f_format_correct = 0
        for row in range(2, 97):
            cell = ws.cell(row=row, column=6)
            # Accept 0.0% or similar percentage formats
            nf = cell.number_format or ''
            if '0.0%' in nf or nf == '0%' or '%' in nf:
                f_format_correct += 1

        ratio = f_format_correct / 95
        if ratio >= 0.95:
            sample_fmt = ws.cell(row=2, column=6).number_format
            print(f"PASS: Component 2 — F column number format correct ({f_format_correct}/95, sample: '{sample_fmt}') (0.20 pts)")
            total_score += 0.20
        else:
            sample_fmt = ws.cell(row=2, column=6).number_format
            print(f"FAIL: Component 2 — F column number format incorrect ({f_format_correct}/95, sample: '{sample_fmt}')")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2:G96 contain formula =DATEDIF(E{row},TODAY(),"Y") (0.30 points)
    # This checks the years since last raise formula
    try:
        g_formula_correct = 0
        g_formula_total = 95  # rows 2 through 96
        for row in range(2, 97):
            cell_val = ws.cell(row=row, column=7).value  # Column G
            if cell_val is None:
                continue
            formula_norm = normalize_formula(str(cell_val))
            expected_norm = normalize_formula(f'=DATEDIF(E{row},TODAY(),"Y")')
            if formula_norm == expected_norm:
                g_formula_correct += 1
            else:
                if g_formula_correct == 0 and row == 2:
                    print(f"FAIL sample: G{row} has {repr(cell_val)}, expected =DATEDIF(E{row},TODAY(),\"Y\")")

        ratio = g_formula_correct / g_formula_total
        if ratio >= 0.95:
            print(f"PASS: Component 3 — G column formulas correct ({g_formula_correct}/{g_formula_total}) (0.30 pts)")
            total_score += 0.30
        elif ratio >= 0.5:
            partial = round(0.15 * ratio / 0.95, 2)
            print(f"PARTIAL: Component 3 — G column formulas partially correct ({g_formula_correct}/{g_formula_total}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — G column formulas largely missing/incorrect ({g_formula_correct}/{g_formula_total})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: G2:G96 formatted as integer (number format '0') (0.10 points)
    # The initial file has 'General' format for G cells — only explicit '0' (integer) format qualifies
    try:
        g_format_correct = 0
        for row in range(2, 97):
            cell = ws.cell(row=row, column=7)
            nf = cell.number_format or ''
            # Only accept explicit integer format '0', NOT 'General' (which is the default/initial state)
            if nf == '0':
                g_format_correct += 1

        ratio = g_format_correct / 95
        if ratio >= 0.95:
            sample_fmt = ws.cell(row=2, column=7).number_format
            print(f"PASS: Component 4 — G column integer format '0' ({g_format_correct}/95, sample: '{sample_fmt}') (0.10 pts)")
            total_score += 0.10
        else:
            sample_fmt = ws.cell(row=2, column=7).number_format
            print(f"FAIL: Component 4 — G column format not '0' integer ({g_format_correct}/95, sample: '{sample_fmt}')")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on A2:G96 with amber (#FFC000) fill when G>=2 (0.10 points)
    try:
        cf_found = False
        amber_found = False
        range_correct = False

        cf_rules = ws.conditional_formatting
        for cf_range in cf_rules:
            cf_range_str = str(cf_range)
            for rule in cf_range.rules:
                # Check if rule uses formula for G>=2 condition
                if rule.type == 'expression' and rule.formula:
                    formula_str = ' '.join(str(f) for f in rule.formula).upper()
                    if 'G' in formula_str and '>=2' in formula_str.replace(' ', '') or '>=' in formula_str:
                        cf_found = True
                        # Check the fill color for amber
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fill = rule.dxf.fill
                            if fill.fgColor:
                                color_rgb = fill.fgColor.rgb
                                # Accept FFFFC000 (amber)
                                if color_rgb and ('FFC000' in color_rgb.upper()):
                                    amber_found = True
                        # Check range covers A2:G96
                        if 'A2' in cf_range_str and 'G96' in cf_range_str:
                            range_correct = True

        if cf_found and amber_found and range_correct:
            print(f"PASS: Component 5 — Conditional formatting with amber fill for G>=2 on A2:G96 (0.10 pts)")
            total_score += 0.10
        elif cf_found and amber_found:
            print(f"PARTIAL: Component 5 — CF amber fill found but range may not be A2:G96 (range: {cf_range_str})")
            total_score += 0.05
        elif cf_found:
            print(f"FAIL: Component 5 — CF found for G>=2 but amber fill (#FFC000) missing")
        else:
            # Try checking if any CF exists at all
            cf_count = sum(1 for _ in cf_rules)
            print(f"FAIL: Component 5 — No conditional formatting found for G>=2 condition (CF ranges: {cf_count})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
