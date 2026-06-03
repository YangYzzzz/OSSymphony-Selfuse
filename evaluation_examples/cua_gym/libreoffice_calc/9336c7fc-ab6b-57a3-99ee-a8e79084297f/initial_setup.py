"""
Initial Setup: Trade Show Booth Preparation and Results Tracker
Task ID: calc_grs_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # =========================================================================
    # Common styles
    # =========================================================================
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    section_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    currency_fmt = '$#,##0.00'
    date_fmt = 'yyyy-mm-dd'

    # =========================================================================
    # Sheet 1: Pre-Show Planning
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Pre-Show Planning"

    headers1 = ["Task Description", "Responsible Person", "Due Date",
                 "Status", "Cost", "Budget vs Actual"]
    col_widths1 = [40, 20, 14, 16, 14, 16]

    for c, (h, w) in enumerate(zip(headers1, col_widths1), 1):
        cell = ws1.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws1.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    ws1.row_dimensions[1].height = 28
    ws1.freeze_panes = "A2"

    # Base date for due dates
    show_date = date(2026, 6, 15)

    # Section data: (section_name, tasks)
    # Each task: (description, person, days_before_show, status, cost, budget_vs_actual)
    sections = [
        ("Booth Design", [
            ("Finalize booth layout and floor plan", "Rachel Kim", 55, "Complete", 2500.00, "Under Budget"),
            ("Design backdrop graphics and signage", "Tom Fischer", 50, "Complete", 4800.00, "On Budget"),
            ("Order custom lighting fixtures", "Rachel Kim", 45, "Complete", 1750.00, "Over Budget"),
            ("Arrange furniture rental (tables, chairs, display stands)", "Lisa Patel", 40, "In Progress", 3200.00, ""),
        ]),
        ("Materials/Supplies", [
            ("Print brochures and product catalogs (500 copies)", "Amy Nguyen", 35, "Complete", 2100.00, "Under Budget"),
            ("Order branded giveaways (pens, tote bags, USB drives)", "Amy Nguyen", 38, "Complete", 3500.00, "On Budget"),
            ("Prepare product demo units and prototypes", "Derek Owens", 30, "In Progress", 5200.00, ""),
            ("Ship materials to venue (freight coordinator)", "Carlos Ruiz", 20, "Not Started", 1800.00, ""),
        ]),
        ("Staffing", [
            ("Confirm booth staff schedule (8 staff, 3 shifts)", "Lisa Patel", 28, "In Progress", 6400.00, ""),
            ("Arrange travel and hotel for out-of-town staff", "Lisa Patel", 35, "Complete", 8500.00, "Over Budget"),
            ("Schedule product demo training sessions", "Derek Owens", 25, "Not Started", 0.00, ""),
            ("Hire temporary booth assistants (2 people)", "Carlos Ruiz", 22, "Not Started", 1600.00, ""),
        ]),
        ("Lead Capture Setup", [
            ("Configure lead capture tablets and app", "Jin Tanaka", 18, "Not Started", 1200.00, ""),
            ("Create lead qualification questionnaire", "Jin Tanaka", 20, "In Progress", 0.00, ""),
            ("Set up CRM integration for real-time sync", "Jin Tanaka", 15, "Not Started", 750.00, ""),
        ]),
        ("Marketing", [
            ("Send pre-show email campaign to prospects", "Samantha Brooks", 21, "Not Started", 350.00, ""),
            ("Schedule social media posts (before/during show)", "Samantha Brooks", 14, "Not Started", 200.00, ""),
            ("Prepare press kit for media attendees", "Tom Fischer", 18, "Not Started", 650.00, ""),
        ]),
    ]

    row = 2
    for section_name, tasks in sections:
        # Section header row (merged across all columns)
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws1.cell(row=row, column=1, value=section_name)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border
        for c in range(2, 7):
            ws1.cell(row=row, column=c).fill = section_fill
            ws1.cell(row=row, column=c).border = thin_border
        row += 1

        for desc, person, days_before, status, cost, budget in tasks:
            due = show_date - timedelta(days=days_before)
            ws1.cell(row=row, column=1, value=desc).border = thin_border
            ws1.cell(row=row, column=2, value=person).border = thin_border
            c_date = ws1.cell(row=row, column=3, value=due)
            c_date.number_format = date_fmt
            c_date.border = thin_border
            c_date.alignment = Alignment(horizontal="center")
            c_status = ws1.cell(row=row, column=4, value=status)
            c_status.border = thin_border
            c_status.alignment = Alignment(horizontal="center")
            c_cost = ws1.cell(row=row, column=5, value=cost)
            c_cost.number_format = currency_fmt
            c_cost.border = thin_border
            c_budget = ws1.cell(row=row, column=6, value=budget)
            c_budget.border = thin_border
            c_budget.alignment = Alignment(horizontal="center")
            row += 1

    last_data_row = row - 1

    # Status dropdown
    status_dv = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,Cancelled"',
        allow_blank=True,
        showDropDown=False,
    )
    status_dv.error = "Please select a valid status"
    status_dv.errorTitle = "Invalid Status"
    status_dv.add(f"D2:D{last_data_row}")
    ws1.add_data_validation(status_dv)

    # Conditional formatting: highlight overdue tasks (Due Date < TODAY and Status != Complete)
    red_fill_style = DifferentialStyle(
        fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
        font=Font(color="9C0006"),
    )
    ws1.conditional_formatting.add(
        f"A2:F{last_data_row}",
        FormulaRule(
            formula=[f'AND($C2<TODAY(),$D2<>"Complete")'],
            fill=PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid"),
        )
    )

    # =========================================================================
    # Sheet 2: Lead Capture
    # =========================================================================
    ws2 = wb.create_sheet("Lead Capture")

    headers2 = ["Lead ID", "First Name", "Last Name", "Company", "Title",
                 "Email", "Phone", "Products Interested In", "Lead Quality",
                 "Notes", "Follow-Up Date", "Assigned To"]
    col_widths2 = [10, 14, 14, 22, 20, 28, 16, 22, 14, 30, 14, 16]

    for c, (h, w) in enumerate(zip(headers2, col_widths2), 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    ws2.row_dimensions[1].height = 28
    ws2.freeze_panes = "A2"

    leads = [
        ("L001", "David", "Martinez", "Pinnacle Solutions", "VP of Operations", "dmartinez@pinnaclesol.com", "(312) 555-0147", "Enterprise Platform", "Hot", "Very interested in Q3 deployment, budget approved", date(2026, 6, 22), "Jin Tanaka"),
        ("L002", "Karen", "Whitfield", "Nexus Technologies", "IT Director", "kwhitfield@nexustech.com", "(415) 555-0283", "Cloud Suite", "Hot", "Current vendor contract ending August, decision by July", date(2026, 6, 20), "Derek Owens"),
        ("L003", "Brian", "Okafor", "Summit Industries", "Procurement Manager", "bokafor@summitind.com", "(214) 555-0391", "Enterprise Platform", "Warm", "Evaluating 3 vendors, needs ROI comparison", date(2026, 6, 25), "Jin Tanaka"),
        ("L004", "Emily", "Sorensen", "Vanguard Corp", "CTO", "esorensen@vanguardcorp.com", "(617) 555-0452", "Analytics Module", "Hot", "Wants live demo for executive team next week", date(2026, 6, 19), "Samantha Brooks"),
        ("L005", "Michael", "Chandra", "Atlas Group", "Business Analyst", "mchandra@atlasgroup.com", "(503) 555-0178", "Cloud Suite", "Warm", "Interested but needs internal approval", date(2026, 6, 28), "Derek Owens"),
        ("L006", "Jessica", "Thornton", "Meridian Consulting", "Partner", "jthornton@meridianconsult.com", "(212) 555-0634", "Enterprise Platform", "Cold", "Just browsing, no immediate need", date(2026, 7, 10), "Jin Tanaka"),
        ("L007", "Robert", "Nakamura", "Vertex Dynamics", "Engineering Lead", "rnakamura@vertexdyn.com", "(408) 555-0729", "Analytics Module", "Warm", "Technical evaluation phase, needs API documentation", date(2026, 6, 24), "Derek Owens"),
        ("L008", "Alicia", "Bergström", "Nordic Systems AB", "Director of Innovation", "abergstrom@nordicsys.se", "(206) 555-0845", "Cloud Suite", "Hot", "International expansion project, high budget", date(2026, 6, 21), "Samantha Brooks"),
        ("L009", "Trevor", "Washington", "Crestline Financial", "VP of Technology", "twashington@crestlinefin.com", "(404) 555-0156", "Enterprise Platform", "Warm", "Regulated industry, needs compliance review", date(2026, 6, 30), "Jin Tanaka"),
        ("L010", "Priya", "Sharma", "Horizon Health", "Clinical Systems Manager", "psharma@horizonhealth.org", "(713) 555-0293", "Analytics Module", "Cold", "Early research stage, 2027 budget cycle", date(2026, 7, 15), "Samantha Brooks"),
        ("L011", "James", "O'Brien", "Cascade Manufacturing", "Plant Manager", "jobrien@cascademfg.com", "(206) 555-0467", "Enterprise Platform", "Warm", "Replacing legacy system, timeline flexible", date(2026, 6, 27), "Derek Owens"),
        ("L012", "Linda", "Vasquez", "Apex Retail Group", "Chief Digital Officer", "lvasquez@apexretail.com", "(310) 555-0581", "Cloud Suite", "Hot", "Multi-location rollout planned, wants pricing for 50+ seats", date(2026, 6, 20), "Jin Tanaka"),
        ("L013", "Steven", "Kowalski", "Ironclad Defense", "Systems Architect", "skowalski@ironcladdef.com", "(571) 555-0312", "Analytics Module", "Cold", "Government contractor, long procurement cycle", date(2026, 7, 12), "Derek Owens"),
        ("L014", "Monica", "Chen-Li", "Pacific Trade Alliance", "Operations Director", "mchenli@pacifictrade.com", "(415) 555-0698", "Enterprise Platform", "Hot", "Expanding to 3 new markets, urgent need", date(2026, 6, 18), "Samantha Brooks"),
        ("L015", "Patrick", "Dubois", "EuroLink Solutions", "Managing Director", "pdubois@eurolink.eu", "(646) 555-0834", "Cloud Suite", "Warm", "European data residency requirements, GDPR focus", date(2026, 6, 26), "Jin Tanaka"),
    ]

    for r, lead in enumerate(leads, 2):
        for c, val in enumerate(lead, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 11:  # Follow-Up Date
                cell.number_format = date_fmt
                cell.alignment = Alignment(horizontal="center")
            elif c == 1:  # Lead ID
                cell.alignment = Alignment(horizontal="center")
            elif c == 9:  # Lead Quality
                cell.alignment = Alignment(horizontal="center")

    # Products dropdown
    products_dv = DataValidation(
        type="list",
        formula1='"Enterprise Platform,Cloud Suite,Analytics Module,Security Add-on,Support Package"',
        allow_blank=True,
        showDropDown=False,
    )
    products_dv.add("H2:H100")
    ws2.add_data_validation(products_dv)

    # Lead Quality dropdown
    quality_dv = DataValidation(
        type="list",
        formula1='"Hot,Warm,Cold"',
        allow_blank=True,
        showDropDown=False,
    )
    quality_dv.add("I2:I100")
    ws2.add_data_validation(quality_dv)

    # =========================================================================
    # Sheet 3: Post-Show Analysis
    # =========================================================================
    ws3 = wb.create_sheet("Post-Show Analysis")

    # Title
    ws3.merge_cells("A1:D1")
    title_cell = ws3.cell(row=1, column=1, value="Post-Show Analysis")
    title_cell.font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 32

    # --- Leads Summary Section ---
    ws3.cell(row=3, column=1, value="Lead Metrics").font = Font(bold=True, size=12, color="2F5496")
    ws3.column_dimensions["A"].width = 28
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 20
    ws3.column_dimensions["D"].width = 18

    labels_leads = [
        ("Total Leads Captured", None),
        ("Hot Leads", None),
        ("Warm Leads", None),
        ("Cold Leads", None),
    ]
    for i, (label, val) in enumerate(labels_leads):
        r = 4 + i
        ws3.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=r, column=1).border = thin_border
        ws3.cell(row=r, column=2).border = thin_border

    # --- Cost Summary Section ---
    ws3.cell(row=9, column=1, value="Cost Analysis").font = Font(bold=True, size=12, color="2F5496")

    cost_labels = [
        ("Total Booth Cost", None),
        ("Cost Per Lead", None),
    ]
    for i, (label, val) in enumerate(cost_labels):
        r = 10 + i
        ws3.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=r, column=1).border = thin_border
        ws3.cell(row=r, column=2).border = thin_border
        ws3.cell(row=r, column=2).number_format = currency_fmt

    # --- Previous Show Comparison Section ---
    ws3.cell(row=13, column=1, value="Comparison to Previous Shows").font = Font(bold=True, size=12, color="2F5496")

    comp_headers = ["Metric", "2024 Show", "2025 Show", "2026 Show"]
    for c, h in enumerate(comp_headers, 1):
        cell = ws3.cell(row=14, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    comp_data = [
        ("Total Leads", 42, 68, None),
        ("Hot Leads", 8, 15, None),
        ("Total Booth Cost", 38000, 42000, None),
        ("Cost Per Lead", None, None, None),
        ("Conversion Rate", "12%", "18%", ""),
    ]
    for i, (metric, v2024, v2025, v2026) in enumerate(comp_data):
        r = 15 + i
        ws3.cell(row=r, column=1, value=metric).border = thin_border
        ws3.cell(row=r, column=1).font = Font(bold=True)
        c2 = ws3.cell(row=r, column=2, value=v2024)
        c2.border = thin_border
        c2.alignment = Alignment(horizontal="center")
        c3 = ws3.cell(row=r, column=3, value=v2025)
        c3.border = thin_border
        c3.alignment = Alignment(horizontal="center")
        c4 = ws3.cell(row=r, column=4, value=v2026)
        c4.border = thin_border
        c4.alignment = Alignment(horizontal="center")
        if metric in ("Total Booth Cost", "Cost Per Lead"):
            c2.number_format = currency_fmt
            c3.number_format = currency_fmt
            c4.number_format = currency_fmt

    # --- Lead Quality Distribution Data (for chart) ---
    ws3.cell(row=21, column=1, value="Lead Quality Distribution").font = Font(bold=True, size=12, color="2F5496")
    ws3.cell(row=22, column=1, value="Quality Level").font = Font(bold=True)
    ws3.cell(row=22, column=1).border = thin_border
    ws3.cell(row=22, column=2, value="Count").font = Font(bold=True)
    ws3.cell(row=22, column=2).border = thin_border

    quality_data = [("Hot", None), ("Warm", None), ("Cold", None)]
    for i, (qual, count) in enumerate(quality_data):
        r = 23 + i
        ws3.cell(row=r, column=1, value=qual).border = thin_border
        ws3.cell(row=r, column=2, value=count).border = thin_border

    # NOTE: No formulas, no charts in initial state — agent must create these

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
