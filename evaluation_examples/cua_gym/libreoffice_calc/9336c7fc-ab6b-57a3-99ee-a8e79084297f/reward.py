"""
Reward Script: Trade Show Booth Preparation and Results Tracker
Task ID: calc_grs_085
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Post-Show Analysis formulas for lead metrics (B4-B7)
  Component 2 (0.25): Post-Show Analysis formulas for cost analysis (B10, B11)
  Component 3 (0.25): Post-Show Analysis comparison formulas linking to current show (D15-D17, D18/B11)
  Component 4 (0.20): Chart exists on Post-Show Analysis sheet (funnel/bar chart for lead quality)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_085'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Key insight: The initial file has the same 3 sheets and same structure,
    but the Post-Show Analysis sheet has NO formulas (all None in column B
    for metrics rows, and no formulas in D15-D18) and NO chart.
    The golden file adds all the formulas and a chart.
    So all scoring components target the Post-Show Analysis formulas and chart.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Must have Post-Show Analysis sheet
    if 'Post-Show Analysis' not in wb.sheetnames:
        print("CRITICAL: 'Post-Show Analysis' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Post-Show Analysis']

    # Component 1: Lead metrics formulas in B4-B7 (0.30 points)
    # B4 = total leads (COUNTA on Lead Capture!A column)
    # B5 = hot leads (COUNTIF on Lead Capture!I column for "Hot")
    # B6 = warm leads (COUNTIF on Lead Capture!I column for "Warm")
    # B7 = cold leads (COUNTIF on Lead Capture!I column for "Cold")
    try:
        formulas_found = 0
        b4 = ws['B4'].value
        b5 = ws['B5'].value
        b6 = ws['B6'].value
        b7 = ws['B7'].value

        # B4: should be a formula referencing Lead Capture for counting leads
        if b4 and isinstance(b4, str) and '=' in b4:
            b4_upper = b4.upper().replace(' ', '')
            if 'COUNTA' in b4_upper or 'COUNT' in b4_upper or 'LEADCAPTURE' in b4_upper.replace("'", "").replace(" ", ""):
                formulas_found += 1
                print(f"PASS: B4 has lead count formula: {b4}")
            else:
                print(f"PARTIAL: B4 has formula but unexpected: {b4}")
        else:
            print(f"FAIL: B4 should have a count formula, found: {b4}")

        # B5: COUNTIF for Hot
        if b5 and isinstance(b5, str) and '=' in b5:
            b5_upper = b5.upper().replace(' ', '')
            if 'COUNTIF' in b5_upper and 'HOT' in b5_upper:
                formulas_found += 1
                print(f"PASS: B5 has Hot leads COUNTIF formula: {b5}")
            else:
                print(f"PARTIAL: B5 has formula but unexpected: {b5}")
        else:
            print(f"FAIL: B5 should have a COUNTIF formula for Hot, found: {b5}")

        # B6: COUNTIF for Warm
        if b6 and isinstance(b6, str) and '=' in b6:
            b6_upper = b6.upper().replace(' ', '')
            if 'COUNTIF' in b6_upper and 'WARM' in b6_upper:
                formulas_found += 1
                print(f"PASS: B6 has Warm leads COUNTIF formula: {b6}")
            else:
                print(f"PARTIAL: B6 has formula but unexpected: {b6}")
        else:
            print(f"FAIL: B6 should have a COUNTIF formula for Warm, found: {b6}")

        # B7: COUNTIF for Cold
        if b7 and isinstance(b7, str) and '=' in b7:
            b7_upper = b7.upper().replace(' ', '')
            if 'COUNTIF' in b7_upper and 'COLD' in b7_upper:
                formulas_found += 1
                print(f"PASS: B7 has Cold leads COUNTIF formula: {b7}")
            else:
                print(f"PARTIAL: B7 has formula but unexpected: {b7}")
        else:
            print(f"FAIL: B7 should have a COUNTIF formula for Cold, found: {b7}")

        component1_score = 0.30 * (formulas_found / 4)
        if formulas_found > 0:
            print(f"PASS: Component 1 - Lead metrics formulas: {formulas_found}/4 ({component1_score:.2f} pts)")
            total_score += component1_score
        else:
            print(f"FAIL: Component 1 - No lead metric formulas found in B4:B7")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Cost analysis formulas B10, B11 (0.25 points)
    # B10 = Total Booth Cost (SUM/SUMPRODUCT from Pre-Show Planning costs)
    # B11 = Cost Per Lead (B10/B4 or similar)
    try:
        cost_formulas = 0
        b10 = ws['B10'].value
        b11 = ws['B11'].value

        # B10: should reference Pre-Show Planning costs
        if b10 and isinstance(b10, str) and '=' in b10:
            b10_upper = b10.upper().replace(' ', '')
            if ('SUM' in b10_upper or 'PRE-SHOWPLANNING' in b10_upper.replace("'", "").replace(" ", "")
                or 'PRE' in b10_upper):
                cost_formulas += 1
                print(f"PASS: B10 has total cost formula: {b10}")
            else:
                print(f"PARTIAL: B10 has formula but unexpected: {b10}")
        else:
            print(f"FAIL: B10 should have a cost sum formula, found: {b10}")

        # B11: cost per lead = cost / leads
        if b11 and isinstance(b11, str) and '=' in b11:
            b11_upper = b11.upper().replace(' ', '')
            if ('B10' in b11_upper or 'B4' in b11_upper or 'COST' in b11_upper
                or '/' in b11_upper):
                cost_formulas += 1
                print(f"PASS: B11 has cost per lead formula: {b11}")
            else:
                print(f"PARTIAL: B11 has formula but unexpected: {b11}")
        else:
            print(f"FAIL: B11 should have a cost per lead formula, found: {b11}")

        component2_score = 0.25 * (cost_formulas / 2)
        if cost_formulas > 0:
            print(f"PASS: Component 2 - Cost analysis formulas: {cost_formulas}/2 ({component2_score:.3f} pts)")
            total_score += component2_score
        else:
            print(f"FAIL: Component 2 - No cost analysis formulas found in B10:B11")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Comparison table formulas linking 2026 show to current data (0.25 points)
    # D15 = =B4 (total leads for 2026)
    # D16 = =B5 (hot leads for 2026)
    # D17 = =B10 (total booth cost for 2026)
    # D18 = =B11 (cost per lead for 2026)
    # Also B23-B25 should reference B5-B7 for chart data
    try:
        comparison_formulas = 0
        cells_to_check = {
            'D15': 'Total Leads 2026',
            'D16': 'Hot Leads 2026',
            'D17': 'Total Booth Cost 2026',
            'D18': 'Cost Per Lead 2026',
        }
        for cell_ref, desc in cells_to_check.items():
            val = ws[cell_ref].value
            if val and isinstance(val, str) and '=' in val:
                comparison_formulas += 1
                print(f"PASS: {cell_ref} has formula for {desc}: {val}")
            else:
                print(f"FAIL: {cell_ref} should have formula for {desc}, found: {val}")

        # Also check B23-B25 (lead quality distribution data for chart)
        chart_data_formulas = 0
        for row_num, label in [(23, 'Hot'), (24, 'Warm'), (25, 'Cold')]:
            val = ws.cell(row=row_num, column=2).value
            if val and isinstance(val, str) and '=' in val:
                chart_data_formulas += 1
                print(f"PASS: B{row_num} has formula for {label} count: {val}")
            else:
                print(f"INFO: B{row_num} for {label} count: {val}")

        # Score based on comparison formulas (main) + partial for chart data
        total_comp3 = comparison_formulas + chart_data_formulas
        max_comp3 = 7  # 4 comparison + 3 chart data
        component3_score = 0.25 * (total_comp3 / max_comp3)
        if total_comp3 > 0:
            print(f"PASS: Component 3 - Comparison/chart-data formulas: {total_comp3}/{max_comp3} ({component3_score:.3f} pts)")
            total_score += component3_score
        else:
            print(f"FAIL: Component 3 - No comparison formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Chart on Post-Show Analysis sheet (0.20 points)
    # Task requires a "funnel chart for lead quality distribution"
    # openpyxl represents funnel as BarChart typically
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_type = type(chart).__name__
            print(f"PASS: Component 4 - Chart found on Post-Show Analysis (type: {chart_type}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 - No chart found on Post-Show Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
