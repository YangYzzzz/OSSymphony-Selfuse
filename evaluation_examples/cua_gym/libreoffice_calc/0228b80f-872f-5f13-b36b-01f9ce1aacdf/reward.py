"""
Reward Script: Merge cells A1:H1 to create a centered title cell
Task ID: calc_gsi_089
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): A1:H1 merge range exists
  Component 2 (0.25): A1 horizontal alignment is 'center'
  Component 3 (0.25): A1 value is 'Annual Performance Review 2024' and B1-H1 are MergedCell
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_089'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: A1:H1 merge range exists (0.50 points)
    # This is the core task requirement: cells A1 through H1 must be merged.
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        # Check that A1:H1 is among the merged ranges
        a1_h1_merged = False
        for mr in ws.merged_cells.ranges:
            # The merge range should cover A1:H1 (min_col=1, max_col=8, min_row=1, max_row=1)
            if mr.min_row == 1 and mr.max_row == 1 and mr.min_col == 1 and mr.max_col == 8:
                a1_h1_merged = True
                break
        if a1_h1_merged:
            print(f"PASS: Component 1 - A1:H1 merge range found (0.50 pts)")
            total_score += 0.50
        else:
            print(f"FAIL: Component 1 - A1:H1 merge range not found. Merged ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: A1 horizontal alignment is 'center' (0.25 points)
    # The task says "Merge and Center Cells", so A1 should be center-aligned.
    try:
        h_align = ws['A1'].alignment.horizontal
        if h_align == 'center':
            print(f"PASS: Component 2 - A1 horizontal alignment is 'center' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 - A1 horizontal alignment is '{h_align}', expected 'center'")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: A1 contains the correct title AND B1-H1 are MergedCell objects (0.25 points)
    # Verifies the title text is preserved and merge physically affected adjacent cells.
    try:
        a1_value = ws['A1'].value
        title_correct = (isinstance(a1_value, str) and
                         a1_value.strip() == 'Annual Performance Review 2024')

        # Check that B1 through H1 are MergedCell (evidence of physical merge)
        all_merged_cells = True
        for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            cell = ws[f'{col_letter}1']
            if not isinstance(cell, MergedCell):
                all_merged_cells = False
                print(f"  DEBUG: {col_letter}1 is {type(cell).__name__}, not MergedCell")
                break

        if title_correct and all_merged_cells:
            print(f"PASS: Component 3 - Title correct and B1:H1 are MergedCell (0.25 pts)")
            total_score += 0.25
        elif title_correct and not all_merged_cells:
            print(f"FAIL: Component 3 - Title correct but B1:H1 are not all MergedCell")
        elif not title_correct:
            print(f"FAIL: Component 3 - A1 value is '{a1_value}', expected 'Annual Performance Review 2024'")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
