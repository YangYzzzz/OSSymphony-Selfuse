"""
Initial Setup: Create a printable performance report with an unmerged title row
Task ID: calc_gsi_089
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_089'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- Row 1: Title (NOT merged - that is the task) ---
    ws["A1"] = "Annual Performance Review 2024"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    # No merge, no center alignment - agent must do that

    # --- Row 2: Headers ---
    headers = ["Employee", "Department", "Q1 Score", "Q2 Score",
               "Q3 Score", "Q4 Score", "Average", "Rating"]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Rows 3-14: Employee data ---
    data = [
        ["Sarah Chen", "Engineering", 87, 91, 89, 93, None, None],
        ["Marcus Johnson", "Marketing", 78, 82, 80, 85, None, None],
        ["Priya Patel", "Finance", 92, 88, 95, 90, None, None],
        ["David Kim", "Engineering", 85, 83, 88, 86, None, None],
        ["Elena Rodriguez", "Sales", 90, 94, 91, 88, None, None],
        ["James O'Brien", "Operations", 76, 79, 81, 84, None, None],
        ["Aisha Mohammed", "HR", 88, 86, 90, 92, None, None],
        ["Robert Zhang", "Finance", 82, 85, 87, 89, None, None],
        ["Maria Santos", "Marketing", 91, 89, 86, 93, None, None],
        ["Thomas Anderson", "Engineering", 84, 88, 90, 87, None, None],
        ["Li Wei", "Sales", 79, 83, 85, 90, None, None],
        ["Catherine Miller", "Operations", 93, 91, 94, 96, None, None],
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal="center", vertical="center")

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c)
            if c == 7:  # Average column - formula
                cell.value = f"=AVERAGE(C{r}:F{r})"
            elif c == 8:  # Rating column - formula
                cell.value = f'=IF(G{r}>=90,"Excellent",IF(G{r}>=85,"Good",IF(G{r}>=80,"Satisfactory","Needs Improvement")))'
            else:
                cell.value = val
            cell.border = data_border
            if c >= 3:
                cell.alignment = center_align

    # Set column widths for readability
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 18

    # Set row 1 height for title
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
