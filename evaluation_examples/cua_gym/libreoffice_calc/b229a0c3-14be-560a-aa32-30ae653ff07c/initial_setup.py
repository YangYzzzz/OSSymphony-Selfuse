"""
Initial Setup: Apply Clone Formatting tool to copy A1 formatting to A10, A20, A30, A40
Task ID: calc_gfl_086
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Template ---
    ws = wb.active
    ws.title = 'Template'

    # Define the section header style for A1
    header_font = Font(name="Calibri", size=14, bold=True, color="00003366")  # dark blue
    header_fill = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")  # light gray
    thick_bottom = Border(
        bottom=Side(style="thick", color="000000")
    )

    # Column headers in row 2
    col_headers = ['Category', 'Item', 'Q1 Budget', 'Q2 Budget', 'Q3 Budget', 'Q4 Budget', 'Annual Total']
    col_widths = [22, 28, 14, 14, 14, 14, 16]

    for c, (h, w) in enumerate(zip(col_headers, col_widths), 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    # A1 - Section header with special formatting
    ws['A1'] = 'Department Budget Overview'
    ws['A1'].font = header_font
    ws['A1'].fill = header_fill
    ws['A1'].border = thick_bottom

    # Data rows 3-9: Personnel section
    personnel_data = [
        ['Personnel', 'Base Salaries', 125000, 125000, 130000, 130000],
        ['Personnel', 'Overtime Pay', 8500, 9200, 7800, 11500],
        ['Personnel', 'Benefits & Insurance', 42000, 42000, 43500, 43500],
        ['Personnel', 'Training & Development', 5000, 3500, 6200, 4800],
        ['Personnel', 'Recruitment Costs', 12000, 8000, 15000, 10000],
        ['Personnel', 'Employee Wellness', 3200, 3200, 3200, 3200],
        ['Personnel', 'Performance Bonuses', 0, 0, 0, 28000],
    ]
    for r, row_data in enumerate(personnel_data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 3:
                cell.number_format = '#,##0'

    # A10 - Sub-section header (NO special formatting - plain text)
    ws['A10'] = 'Technology & Infrastructure'

    # Data rows 11-19
    tech_data = [
        ['Technology', 'Software Licenses', 18500, 18500, 22000, 22000],
        ['Technology', 'Cloud Services (AWS)', 9800, 10200, 10600, 11000],
        ['Technology', 'Hardware Refresh', 25000, 5000, 25000, 5000],
        ['Technology', 'Cybersecurity Tools', 7500, 7500, 8200, 8200],
        ['Technology', 'Network Maintenance', 4200, 4200, 4200, 4200],
        ['Technology', 'IT Support Contracts', 6000, 6000, 6000, 6000],
        ['Technology', 'Data Backup Services', 2800, 2800, 2800, 2800],
        ['Technology', 'Dev Tools & APIs', 3500, 3500, 4000, 4000],
        ['Technology', 'Telecom & Internet', 2200, 2200, 2200, 2200],
    ]
    for r, row_data in enumerate(tech_data, 11):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 3:
                cell.number_format = '#,##0'

    # A20 - Sub-section header (NO special formatting)
    ws['A20'] = 'Marketing & Outreach'

    # Data rows 21-29
    marketing_data = [
        ['Marketing', 'Digital Advertising', 15000, 18000, 20000, 25000],
        ['Marketing', 'Print & Media', 5000, 5000, 3000, 8000],
        ['Marketing', 'Trade Shows & Events', 12000, 0, 12000, 0],
        ['Marketing', 'Content Creation', 4500, 4500, 5000, 5000],
        ['Marketing', 'SEO & Analytics', 3000, 3000, 3500, 3500],
        ['Marketing', 'Social Media Mgmt', 2500, 2500, 2500, 2500],
        ['Marketing', 'Brand Collateral', 3800, 1500, 3800, 1500],
        ['Marketing', 'PR & Communications', 6000, 6000, 6000, 6000],
        ['Marketing', 'Market Research', 8000, 4000, 8000, 4000],
    ]
    for r, row_data in enumerate(marketing_data, 21):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 3:
                cell.number_format = '#,##0'

    # A30 - Sub-section header (NO special formatting)
    ws['A30'] = 'Operations & Facilities'

    # Data rows 31-39
    ops_data = [
        ['Operations', 'Office Rent', 35000, 35000, 35000, 35000],
        ['Operations', 'Utilities', 4800, 4200, 5500, 5000],
        ['Operations', 'Office Supplies', 2200, 2200, 2500, 2500],
        ['Operations', 'Cleaning Services', 1800, 1800, 1800, 1800],
        ['Operations', 'Equipment Leasing', 6500, 6500, 6500, 6500],
        ['Operations', 'Insurance Premiums', 8000, 8000, 8000, 8000],
        ['Operations', 'Legal & Compliance', 5000, 3000, 5000, 7000],
        ['Operations', 'Travel & Transport', 7500, 9000, 8500, 6000],
        ['Operations', 'Miscellaneous', 2000, 2000, 2000, 2000],
    ]
    for r, row_data in enumerate(ops_data, 31):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 3:
                cell.number_format = '#,##0'

    # A40 - Sub-section header (NO special formatting)
    ws['A40'] = 'Strategic Initiatives'

    # Data rows 41-45
    strat_data = [
        ['Strategic', 'Product R&D', 30000, 30000, 35000, 35000],
        ['Strategic', 'Partnership Dev', 8000, 8000, 10000, 10000],
        ['Strategic', 'Expansion Planning', 5000, 5000, 12000, 12000],
        ['Strategic', 'Innovation Lab', 10000, 10000, 10000, 10000],
        ['Strategic', 'Contingency Fund', 15000, 15000, 15000, 15000],
    ]
    for r, row_data in enumerate(strat_data, 41):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 3:
                cell.number_format = '#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
