"""
Initial Setup: Create expense workbook with red tab on Expenses sheet only
Task ID: calc_gsi_072
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

RED_TAB = "FF0000"


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Header style ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0.00'
    date_fmt = 'yyyy-mm-dd'

    # === Sheet 1: Expenses (RED TAB - already color-coded) ===
    ws_expenses = wb.active
    ws_expenses.title = "Expenses"
    ws_expenses.sheet_properties.tabColor = RED_TAB

    headers_exp = ["Date", "Category", "Description", "Amount", "Approved By"]
    for c, h in enumerate(headers_exp, 1):
        cell = ws_expenses.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    expenses_data = [
        ["2025-09-01", "Office", "Printer cartridges", 245.99, "Diana Ross"],
        ["2025-09-03", "Travel", "Flight to Chicago", 389.50, "Mark Stevens"],
        ["2025-09-05", "Meals", "Client lunch - Riviera Grill", 127.80, "Diana Ross"],
        ["2025-09-07", "Supplies", "Whiteboard markers (bulk)", 34.50, "Sarah Chen"],
        ["2025-09-10", "Equipment", "Ergonomic keyboard x3", 297.00, "Mark Stevens"],
        ["2025-09-12", "Travel", "Taxi to airport", 52.00, "Diana Ross"],
        ["2025-09-14", "Other", "Team building event", 450.00, "Sarah Chen"],
        ["2025-09-16", "Meals", "Department pizza day", 89.60, "Mark Stevens"],
        ["2025-09-18", "Supplies", "Sticky notes and binders", 28.75, "Diana Ross"],
        ["2025-09-20", "Equipment", "Monitor stand x2", 159.98, "Sarah Chen"],
        ["2025-09-22", "Travel", "Hotel - 2 nights Denver", 312.00, "Mark Stevens"],
        ["2025-09-25", "Other", "Software license renewal", 599.00, "Diana Ross"],
    ]
    for r, row_data in enumerate(expenses_data, 2):
        ws_expenses.cell(row=r, column=1, value=row_data[0]).number_format = date_fmt
        ws_expenses.cell(row=r, column=2, value=row_data[1])
        ws_expenses.cell(row=r, column=3, value=row_data[2])
        ws_expenses.cell(row=r, column=4, value=row_data[3]).number_format = currency_fmt
        ws_expenses.cell(row=r, column=5, value=row_data[4])

    ws_expenses.column_dimensions["A"].width = 14
    ws_expenses.column_dimensions["B"].width = 12
    ws_expenses.column_dimensions["C"].width = 32
    ws_expenses.column_dimensions["D"].width = 14
    ws_expenses.column_dimensions["E"].width = 16

    # === Sheet 2: Travel (NO tab color) ===
    ws_travel = wb.create_sheet("Travel")
    travel_headers = ["Date", "Destination", "Purpose", "Cost", "Reimbursed"]
    for c, h in enumerate(travel_headers, 1):
        cell = ws_travel.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    travel_data = [
        ["2025-08-05", "New York", "Client meeting", 1245.00, "Yes"],
        ["2025-08-12", "Boston", "Conference", 890.50, "Yes"],
        ["2025-08-20", "San Francisco", "Site visit", 1560.00, "Pending"],
        ["2025-09-01", "Chicago", "Sales pitch", 720.00, "Yes"],
        ["2025-09-08", "Denver", "Training session", 650.00, "No"],
        ["2025-09-15", "Seattle", "Partner meeting", 980.00, "Yes"],
        ["2025-09-22", "Austin", "Tech summit", 1100.00, "Pending"],
        ["2025-10-01", "Miami", "Annual review", 875.00, "No"],
        ["2025-10-10", "Portland", "Workshop", 540.00, "Yes"],
        ["2025-10-18", "Dallas", "Vendor negotiation", 690.00, "Pending"],
    ]
    for r, row_data in enumerate(travel_data, 2):
        ws_travel.cell(row=r, column=1, value=row_data[0]).number_format = date_fmt
        ws_travel.cell(row=r, column=2, value=row_data[1])
        ws_travel.cell(row=r, column=3, value=row_data[2])
        ws_travel.cell(row=r, column=4, value=row_data[3]).number_format = currency_fmt
        ws_travel.cell(row=r, column=5, value=row_data[4])

    ws_travel.column_dimensions["A"].width = 14
    ws_travel.column_dimensions["B"].width = 18
    ws_travel.column_dimensions["C"].width = 24
    ws_travel.column_dimensions["D"].width = 14
    ws_travel.column_dimensions["E"].width = 14

    # === Sheet 3: Meals (NO tab color) ===
    ws_meals = wb.create_sheet("Meals")
    meals_headers = ["Date", "Restaurant", "Attendees", "Amount", "Receipt"]
    for c, h in enumerate(meals_headers, 1):
        cell = ws_meals.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    meals_data = [
        ["2025-08-02", "The Capital Grille", 4, 312.50, "Yes"],
        ["2025-08-09", "Olive Garden", 6, 189.40, "Yes"],
        ["2025-08-15", "Nobu", 3, 425.00, "Yes"],
        ["2025-08-23", "Chipotle (catering)", 12, 156.80, "Yes"],
        ["2025-09-01", "Ruth's Chris", 2, 278.90, "Pending"],
        ["2025-09-10", "Panera Bread", 8, 98.50, "Yes"],
        ["2025-09-17", "Riviera Grill", 3, 210.00, "Yes"],
        ["2025-09-25", "Local Deli", 5, 67.30, "No"],
        ["2025-10-02", "Sushi Zen", 4, 345.60, "Yes"],
        ["2025-10-12", "Pizza Palace", 10, 142.00, "Yes"],
    ]
    for r, row_data in enumerate(meals_data, 2):
        ws_meals.cell(row=r, column=1, value=row_data[0]).number_format = date_fmt
        ws_meals.cell(row=r, column=2, value=row_data[1])
        ws_meals.cell(row=r, column=3, value=row_data[2])
        ws_meals.cell(row=r, column=4, value=row_data[3]).number_format = currency_fmt
        ws_meals.cell(row=r, column=5, value=row_data[4])

    ws_meals.column_dimensions["A"].width = 14
    ws_meals.column_dimensions["B"].width = 22
    ws_meals.column_dimensions["C"].width = 12
    ws_meals.column_dimensions["D"].width = 14
    ws_meals.column_dimensions["E"].width = 10

    # === Sheet 4: Supplies (NO tab color) ===
    ws_supplies = wb.create_sheet("Supplies")
    supplies_headers = ["Date", "Item", "Vendor", "Quantity", "Unit Cost", "Total"]
    for c, h in enumerate(supplies_headers, 1):
        cell = ws_supplies.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    supplies_data = [
        ["2025-08-01", "Copy paper (A4)", "Staples", 20, 8.99, 179.80],
        ["2025-08-08", "Ink cartridges", "Amazon", 5, 32.50, 162.50],
        ["2025-08-14", "Sticky notes", "Office Depot", 30, 2.49, 74.70],
        ["2025-08-22", "Binders (3-ring)", "Staples", 15, 4.99, 74.85],
        ["2025-09-03", "Whiteboard markers", "Amazon", 24, 1.75, 42.00],
        ["2025-09-11", "Envelopes (#10)", "Office Depot", 500, 0.05, 25.00],
        ["2025-09-19", "Pens (ballpoint)", "Staples", 100, 0.35, 35.00],
        ["2025-09-28", "Notebooks (spiral)", "Amazon", 25, 3.99, 99.75],
        ["2025-10-05", "File folders", "Office Depot", 50, 0.89, 44.50],
        ["2025-10-15", "Paper clips (box)", "Staples", 10, 2.99, 29.90],
    ]
    for r, row_data in enumerate(supplies_data, 2):
        ws_supplies.cell(row=r, column=1, value=row_data[0]).number_format = date_fmt
        ws_supplies.cell(row=r, column=2, value=row_data[1])
        ws_supplies.cell(row=r, column=3, value=row_data[2])
        ws_supplies.cell(row=r, column=4, value=row_data[3])
        ws_supplies.cell(row=r, column=5, value=row_data[4]).number_format = currency_fmt
        ws_supplies.cell(row=r, column=6, value=row_data[5]).number_format = currency_fmt

    ws_supplies.column_dimensions["A"].width = 14
    ws_supplies.column_dimensions["B"].width = 22
    ws_supplies.column_dimensions["C"].width = 16
    ws_supplies.column_dimensions["D"].width = 10
    ws_supplies.column_dimensions["E"].width = 12
    ws_supplies.column_dimensions["F"].width = 12

    # === Sheet 5: Equipment (NO tab color) ===
    ws_equip = wb.create_sheet("Equipment")
    equip_headers = ["Date", "Item", "Brand", "Cost", "Department", "Asset Tag"]
    for c, h in enumerate(equip_headers, 1):
        cell = ws_equip.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    equip_data = [
        ["2025-07-15", "Laptop", "Dell XPS 15", 1899.00, "Engineering", "EQ-2025-001"],
        ["2025-07-20", "Monitor 27\"", "LG UltraFine", 549.00, "Design", "EQ-2025-002"],
        ["2025-08-01", "Standing desk", "Uplift V2", 729.00, "Marketing", "EQ-2025-003"],
        ["2025-08-10", "Webcam", "Logitech Brio", 199.99, "HR", "EQ-2025-004"],
        ["2025-08-18", "Keyboard", "Keychron Q1", 169.00, "Engineering", "EQ-2025-005"],
        ["2025-09-02", "Headset", "Jabra Evolve2", 249.00, "Sales", "EQ-2025-006"],
        ["2025-09-12", "Docking station", "CalDigit TS4", 399.99, "Engineering", "EQ-2025-007"],
        ["2025-09-20", "Mouse", "Logitech MX Master", 99.99, "Finance", "EQ-2025-008"],
        ["2025-10-01", "Printer", "HP LaserJet Pro", 449.00, "Admin", "EQ-2025-009"],
        ["2025-10-08", "UPS Battery", "APC 1500VA", 279.00, "IT", "EQ-2025-010"],
    ]
    for r, row_data in enumerate(equip_data, 2):
        ws_equip.cell(row=r, column=1, value=row_data[0]).number_format = date_fmt
        ws_equip.cell(row=r, column=2, value=row_data[1])
        ws_equip.cell(row=r, column=3, value=row_data[2])
        ws_equip.cell(row=r, column=4, value=row_data[3]).number_format = currency_fmt
        ws_equip.cell(row=r, column=5, value=row_data[4])
        ws_equip.cell(row=r, column=6, value=row_data[5])

    ws_equip.column_dimensions["A"].width = 14
    ws_equip.column_dimensions["B"].width = 18
    ws_equip.column_dimensions["C"].width = 20
    ws_equip.column_dimensions["D"].width = 14
    ws_equip.column_dimensions["E"].width = 14
    ws_equip.column_dimensions["F"].width = 16

    # === Sheet 6: Other (NO tab color) ===
    ws_other = wb.create_sheet("Other")
    other_headers = ["Date", "Description", "Category", "Amount", "Notes"]
    for c, h in enumerate(other_headers, 1):
        cell = ws_other.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    other_data = [
        ["2025-08-05", "Team building event", "Morale", 450.00, "Escape room for 12 people"],
        ["2025-08-12", "Software license (Figma)", "Tools", 180.00, "Annual renewal"],
        ["2025-08-19", "Professional development book", "Learning", 45.99, "Clean Architecture"],
        ["2025-09-01", "Subscription (Slack Pro)", "Tools", 599.00, "12-month plan"],
        ["2025-09-09", "Conference registration", "Learning", 350.00, "DevOps Days 2025"],
        ["2025-09-16", "Parking passes (monthly)", "Facilities", 240.00, "3 passes"],
        ["2025-09-24", "Holiday decorations", "Morale", 85.50, "Fall decorations"],
        ["2025-10-01", "Charity donation", "Community", 500.00, "Company match program"],
        ["2025-10-10", "Plant service", "Facilities", 120.00, "Quarterly maintenance"],
        ["2025-10-20", "AWS hosting credits", "Tools", 750.00, "Q4 allocation"],
    ]
    for r, row_data in enumerate(other_data, 2):
        ws_other.cell(row=r, column=1, value=row_data[0]).number_format = date_fmt
        ws_other.cell(row=r, column=2, value=row_data[1])
        ws_other.cell(row=r, column=3, value=row_data[2])
        ws_other.cell(row=r, column=4, value=row_data[3]).number_format = currency_fmt
        ws_other.cell(row=r, column=5, value=row_data[4])

    ws_other.column_dimensions["A"].width = 14
    ws_other.column_dimensions["B"].width = 30
    ws_other.column_dimensions["C"].width = 14
    ws_other.column_dimensions["D"].width = 14
    ws_other.column_dimensions["E"].width = 28

    # === Sheet 7: Summary (NO tab color - not in expense group) ===
    ws_summary = wb.create_sheet("Summary")
    summary_headers = ["Category", "Total Budget", "Spent YTD", "Remaining"]
    for c, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    summary_data = [
        ["Travel", 15000.00, 9251.50, 5748.50],
        ["Meals", 5000.00, 2226.00, 2774.00],
        ["Supplies", 3000.00, 768.00, 2232.00],
        ["Equipment", 25000.00, 5022.97, 19977.03],
        ["Other", 8000.00, 3320.49, 4679.51],
    ]
    for r, row_data in enumerate(summary_data, 2):
        ws_summary.cell(row=r, column=1, value=row_data[0])
        ws_summary.cell(row=r, column=2, value=row_data[1]).number_format = currency_fmt
        ws_summary.cell(row=r, column=3, value=row_data[2]).number_format = currency_fmt
        ws_summary.cell(row=r, column=4, value=row_data[3]).number_format = currency_fmt

    ws_summary.column_dimensions["A"].width = 14
    ws_summary.column_dimensions["B"].width = 16
    ws_summary.column_dimensions["C"].width = 14
    ws_summary.column_dimensions["D"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
