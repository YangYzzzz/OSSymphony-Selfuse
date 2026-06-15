"""
Initial Setup: Investment return calculator - Portfolio sheet with stock data
Task ID: calc_gen_financialformulas_064
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_financialformulas_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Portfolio'

    # --- Headers (Row 1) ---
    headers = ['Ticker', 'Shares', 'Buy Price', 'Current Price',
               'Position Value', 'Return %', 'Weight', 'Weighted Return']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center')

    # --- Stock Data (Rows 2-11): A-D populated, E-H empty ---
    # Realistic portfolio of 10 stocks: Ticker, Shares, Buy Price, Current Price
    stocks = [
        ('AAPL',  150, 142.35, 178.20),
        ('MSFT',  200, 285.60, 415.30),
        ('GOOGL',  50, 2701.50, 3125.80),
        ('AMZN',  100, 3342.00, 3890.75),
        ('NVDA',  300, 410.25, 875.40),
        ('TSLA',   80, 195.80, 248.65),
        ('META',  120, 275.45, 502.30),
        ('BRK.B', 250,  330.20, 418.90),
        ('JNJ',   180,  152.75, 161.40),
        ('JPM',   220,  140.60, 197.85),
    ]

    for r, (ticker, shares, buy_price, cur_price) in enumerate(stocks, 2):
        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=shares)
        ws.cell(row=r, column=3, value=buy_price)
        ws.cell(row=r, column=4, value=cur_price)
        # Columns E (5), F (6), G (7), H (8) are intentionally left empty

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
