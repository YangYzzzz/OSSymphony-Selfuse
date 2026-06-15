"""
Reward Script: Build an investment return calculator with formulas and scatter chart.
Task ID: calc_gen_financialformulas_064
Domain: libreoffice_calc

Scoring:
  Component 1: Position Value formulas (E2:E11) = 0.20 pts
  Component 2: Return % formulas (F2:F11) = 0.20 pts
  Component 3: Weight formulas (G2:G11) = 0.20 pts
  Component 4: Weighted Return formulas (H2:H11) = 0.10 pts
  Component 5: Summary metrics (E13:F15 labels + formulas) = 0.15 pts
  Component 6: Scatter chart present = 0.15 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_financialformulas_064'


def normalize_formula(f):
    """Strip whitespace, uppercase for comparison."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def check_position_value_formulas(ws):
    """
    Component 1: E2:E11 should contain =Bn*Dn (position value = shares * current price).
    Checks that all 10 cells have a multiplication formula referencing B and D columns.
    """
    score = 0.0
    passing = 0
    for row in range(2, 12):
        val = ws.cell(row=row, column=5).value  # column E
        if isinstance(val, str):
            norm = normalize_formula(val)
            # Should match pattern like =B2*D2
            col_letter = 'BCDEFGHIJKLMNOPQRSTUVWXYZ'[row - 2]  # not correct approach
            # Use row-based check: must reference Bn*Dn or Dn*Bn
            b_ref = f'B{row}'
            d_ref = f'D{row}'
            if (b_ref in norm or f'$B${row}' in norm or f'$B$' in norm.replace(str(row), '')) and \
               (d_ref in norm or f'$D${row}' in norm):
                if '*' in norm:
                    passing += 1
    if passing == 10:
        score = 0.20
        print(f"PASS: Component 1 — All 10 Position Value formulas correct in E2:E11 (0.20 pts)")
    elif passing >= 5:
        partial = round(0.10 * passing / 10 * 2, 2)
        score = min(partial, 0.15)
        print(f"PARTIAL: Component 1 — {passing}/10 Position Value formulas correct (score={score})")
    else:
        print(f"FAIL: Component 1 — Only {passing}/10 Position Value formulas present in E2:E11")
    return score


def check_return_pct_formulas(ws):
    """
    Component 2: F2:F11 should contain =(Dn-Cn)/Cn (return %).
    Checks formula structure references D and C columns with subtraction/division.
    """
    score = 0.0
    passing = 0
    for row in range(2, 12):
        val = ws.cell(row=row, column=6).value  # column F
        if isinstance(val, str):
            norm = normalize_formula(val)
            c_ref = f'C{row}'
            d_ref = f'D{row}'
            # Must reference D, C and use both - (subtraction) and / (division)
            if (d_ref in norm or f'$D${row}' in norm) and \
               (c_ref in norm or f'$C${row}' in norm) and \
               '-' in norm and '/' in norm:
                passing += 1
    if passing == 10:
        score = 0.20
        print(f"PASS: Component 2 — All 10 Return % formulas correct in F2:F11 (0.20 pts)")
    elif passing >= 5:
        partial = round(0.10 * passing / 10 * 2, 2)
        score = min(partial, 0.15)
        print(f"PARTIAL: Component 2 — {passing}/10 Return % formulas correct (score={score})")
    else:
        print(f"FAIL: Component 2 — Only {passing}/10 Return % formulas present in F2:F11")
    return score


def check_weight_formulas(ws):
    """
    Component 3: G2:G11 should contain =E2/SUM($E$2:$E$11) (portfolio weight).
    Checks formula references E column and SUM of the full E range (rows 2-11).
    """
    score = 0.0
    passing = 0
    for row in range(2, 12):
        val = ws.cell(row=row, column=7).value  # column G
        if isinstance(val, str):
            norm = normalize_formula(val)
            e_ref = f'E{row}'
            e_ref_dollar = f'$E${row}'
            # Must reference this row's E cell, use SUM, use division,
            # and reference the E2:E11 range (with or without dollar signs)
            has_e_ref = (e_ref in norm or e_ref_dollar in norm)
            has_sum = 'SUM' in norm
            has_div = '/' in norm
            # Check SUM range covers E column rows 2 to 11
            # Accept $E$2:$E$11, E2:E11, $E2:E$11, etc.
            has_e_range = (('$E$2' in norm and '$E$11' in norm) or
                           ('E2:E11' in norm) or
                           ('$E$2:$E$11' in norm))
            if has_e_ref and has_sum and has_div and has_e_range:
                passing += 1
    if passing == 10:
        score = 0.20
        print(f"PASS: Component 3 — All 10 Weight formulas correct in G2:G11 (0.20 pts)")
    elif passing >= 5:
        partial = round(0.10 * passing / 10 * 2, 2)
        score = min(partial, 0.15)
        print(f"PARTIAL: Component 3 — {passing}/10 Weight formulas correct (score={score})")
    else:
        print(f"FAIL: Component 3 — Only {passing}/10 Weight formulas present in G2:G11")
    return score


def check_weighted_return_formulas(ws):
    """
    Component 4: H2:H11 should contain =F2*G2 (weighted return contribution).
    Checks that each cell multiplies Return % (F) by Weight (G).
    """
    score = 0.0
    passing = 0
    for row in range(2, 12):
        val = ws.cell(row=row, column=8).value  # column H
        if isinstance(val, str):
            norm = normalize_formula(val)
            f_ref = f'F{row}'
            g_ref = f'G{row}'
            if (f_ref in norm or f'$F${row}' in norm) and \
               (g_ref in norm or f'$G${row}' in norm) and \
               '*' in norm:
                passing += 1
    if passing == 10:
        score = 0.10
        print(f"PASS: Component 4 — All 10 Weighted Return formulas correct in H2:H11 (0.10 pts)")
    elif passing >= 5:
        partial = round(0.05 * passing / 10 * 2, 2)
        score = min(partial, 0.08)
        print(f"PARTIAL: Component 4 — {passing}/10 Weighted Return formulas correct (score={score})")
    else:
        print(f"FAIL: Component 4 — Only {passing}/10 Weighted Return formulas present in H2:H11")
    return score


def check_summary_metrics(ws):
    """
    Component 5: E13:F15 should have summary labels and formulas.
    - E13: 'Portfolio Return:' (or similar), F13: =SUM(H2:H11)
    - E14: 'Return Std Dev:' (or similar), F14: =STDEV(F2:F11)
    - E15: 'Variance:' (or similar), F15: =VAR(F2:F11)
    """
    score = 0.0
    sub_score = 0.0

    # Check F13: portfolio return = SUM of weighted returns
    try:
        e13 = ws.cell(row=13, column=5).value
        f13 = ws.cell(row=13, column=6).value
        if f13 is not None and isinstance(f13, str) and 'SUM' in normalize_formula(f13) and 'H' in f13.upper():
            sub_score += 0.05
            print(f"PASS: Component 5a — F13 has SUM formula for portfolio return: {repr(f13)}")
        else:
            print(f"FAIL: Component 5a — F13 expected SUM(H...) formula, found: {repr(f13)}")
    except Exception as e:
        print(f"ERROR: Component 5a — {e}")

    # Check F14: std dev of returns
    try:
        e14 = ws.cell(row=14, column=5).value
        f14 = ws.cell(row=14, column=6).value
        if f14 is not None and isinstance(f14, str) and 'STDEV' in normalize_formula(f14):
            sub_score += 0.05
            print(f"PASS: Component 5b — F14 has STDEV formula: {repr(f14)}")
        else:
            print(f"FAIL: Component 5b — F14 expected STDEV formula, found: {repr(f14)}")
    except Exception as e:
        print(f"ERROR: Component 5b — {e}")

    # Check F15: variance
    try:
        e15 = ws.cell(row=15, column=5).value
        f15 = ws.cell(row=15, column=6).value
        if f15 is not None and isinstance(f15, str) and 'VAR' in normalize_formula(f15):
            sub_score += 0.05
            print(f"PASS: Component 5c — F15 has VAR formula: {repr(f15)}")
        else:
            print(f"FAIL: Component 5c — F15 expected VAR formula, found: {repr(f15)}")
    except Exception as e:
        print(f"ERROR: Component 5c — {e}")

    score = sub_score
    if score >= 0.14:
        print(f"PASS: Component 5 — All summary metrics present (0.15 pts)")
    elif score > 0:
        print(f"PARTIAL: Component 5 — Some summary metrics present ({score}/0.15 pts)")
    else:
        print(f"FAIL: Component 5 — No summary metrics found in E13:F15")
    return score


def check_scatter_chart(ws):
    """
    Component 6: A scatter chart should exist plotting Return % vs Position Value.
    Checks that at least one ScatterChart exists with series referencing F and E columns.
    """
    score = 0.0
    try:
        from openpyxl.chart import ScatterChart
        charts = ws._charts
        if not charts:
            print(f"FAIL: Component 6 — No charts found on Portfolio sheet")
            return 0.0

        scatter_found = False
        for chart in charts:
            if isinstance(chart, ScatterChart):
                scatter_found = True
                # Check series references F (Return %) and E (Position Value) columns
                if len(chart.series) >= 1:
                    series = chart.series[0]
                    x_ref = str(series.xVal) if series.xVal else ''
                    y_ref = str(series.yVal) if series.yVal else ''
                    # Check x-axis data references column F and y-axis references column E
                    # or vice versa — either orientation may be acceptable
                    x_has_f = '$F$' in x_ref.upper() or "'F'" in x_ref.upper() or 'F$' in x_ref.upper()
                    y_has_e = '$E$' in y_ref.upper() or "'E'" in y_ref.upper() or 'E$' in y_ref.upper()
                    if x_has_f and y_has_e:
                        score = 0.15
                        print(f"PASS: Component 6 — ScatterChart with Return%(F) vs Position Value(E) found (0.15 pts)")
                    else:
                        # Chart exists but may reference different columns - partial credit
                        score = 0.08
                        print(f"PARTIAL: Component 6 — ScatterChart exists but series references may differ")
                        print(f"  x_ref: {x_ref[:100]}")
                        print(f"  y_ref: {y_ref[:100]}")
                    break

        if not scatter_found:
            # A non-scatter chart exists — partial credit since a chart was created
            print(f"PARTIAL: Component 6 — Chart exists but is not a ScatterChart (0.08 pts)")
            score = 0.08

    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    return score


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Portfolio sheet must exist
    if 'Portfolio' not in wb.sheetnames:
        print(f"CRITICAL: 'Portfolio' sheet not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Portfolio']

    # Precondition: Initial data rows must exist (A2:D11 with tickers)
    ticker = ws.cell(row=2, column=1).value
    if not ticker:
        print("CRITICAL: Initial data appears to be missing (A2 is empty)")
        print("REWARD: 0.0")
        return 0.0

    print(f"File loaded: {file_path}")
    print(f"Sheet: Portfolio, rows={ws.max_row}, cols={ws.max_column}, charts={len(ws._charts)}")
    print()

    # Component 1: Position Value formulas E2:E11 (0.20 pts)
    try:
        total_score += check_position_value_formulas(ws)
    except Exception as e:
        print(f"ERROR: Component 1 check failed: {e}")

    # Component 2: Return % formulas F2:F11 (0.20 pts)
    try:
        total_score += check_return_pct_formulas(ws)
    except Exception as e:
        print(f"ERROR: Component 2 check failed: {e}")

    # Component 3: Weight formulas G2:G11 (0.20 pts)
    try:
        total_score += check_weight_formulas(ws)
    except Exception as e:
        print(f"ERROR: Component 3 check failed: {e}")

    # Component 4: Weighted Return formulas H2:H11 (0.10 pts)
    try:
        total_score += check_weighted_return_formulas(ws)
    except Exception as e:
        print(f"ERROR: Component 4 check failed: {e}")

    # Component 5: Summary metrics E13:F15 (0.15 pts)
    try:
        total_score += check_summary_metrics(ws)
    except Exception as e:
        print(f"ERROR: Component 5 check failed: {e}")

    # Component 6: Scatter chart (0.15 pts)
    try:
        total_score += check_scatter_chart(ws)
    except Exception as e:
        print(f"ERROR: Component 6 check failed: {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
