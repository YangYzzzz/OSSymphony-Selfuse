"""
Initial Setup: Change decimal separator from period to comma in LibreOffice Calc
Task ID: osworld_calc_decimal_separator_001
Domain: libreoffice_calc

Creates a spreadsheet with numeric financial data using English locale (period as decimal
separator). The LibreOffice settings are left at default (period as decimal separator).
The agent must change the decimal separator to comma in LibreOffice Calc settings.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_decimal_separator_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Sales Data ---
    ws1 = wb.active
    ws1.title = 'Sales Data'

    # Headers
    headers = ['Product', 'Region', 'Q1 Revenue', 'Q2 Revenue', 'Q3 Revenue', 'Q4 Revenue', 'Annual Total']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Realistic sales data (using float values with period decimal in Python)
    data = [
        ['Laptop Pro 15',    'North',  45230.75, 52110.50, 48900.25, 61340.00, 207581.50],
        ['Laptop Pro 15',    'South',  38750.00, 41200.75, 36890.50, 49120.25, 165961.50],
        ['Tablet Ultra',     'North',  22400.50, 25600.00, 23100.75, 31500.50, 102601.75],
        ['Tablet Ultra',     'South',  18900.25, 21300.50, 19800.00, 26700.75,  86701.50],
        ['Smartphone X12',   'North',  67800.00, 72400.50, 69500.25, 84200.00, 293900.75],
        ['Smartphone X12',   'South',  54300.75, 58900.00, 55600.50, 71100.25, 239901.50],
        ['Wireless Earbuds', 'North',  12500.50, 14200.25, 13400.00, 18900.75,  59001.50],
        ['Wireless Earbuds', 'South',   9800.00, 11500.75, 10200.50, 14700.25,  46201.50],
        ['Smart Watch S3',   'North',  31200.25, 34500.00, 32800.75, 41600.50, 140101.50],
        ['Smart Watch S3',   'South',  25400.50, 28100.25, 26700.00, 34200.75, 114401.50],
        ['USB-C Hub Pro',    'North',   8900.75,  9800.50,  9200.25, 12400.00,  40301.50],
        ['USB-C Hub Pro',    'South',   7100.00,  8200.75,  7800.50, 10100.25,  33201.50],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c >= 3:
                cell.number_format = '#,##0.00'

    # Column widths
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 10
    for col in ['C', 'D', 'E', 'F', 'G']:
        ws1.column_dimensions[col].width = 14

    # Freeze header row
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Exchange Rates ---
    ws2 = wb.create_sheet('Exchange Rates')

    headers2 = ['Currency', 'Code', 'Rate to USD', 'Rate to EUR', 'Last Updated']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    exchange_data = [
        ['Euro',            'EUR', 0.9231, 1.0000, '2025-06-15'],
        ['British Pound',   'GBP', 0.7854, 0.8509, '2025-06-15'],
        ['Japanese Yen',    'JPY', 149.230, 161.660, '2025-06-15'],
        ['Swiss Franc',     'CHF', 0.8967, 0.9713, '2025-06-15'],
        ['Canadian Dollar', 'CAD', 1.3542, 1.4668, '2025-06-15'],
        ['Australian Dollar','AUD', 1.5234, 1.6501, '2025-06-15'],
        ['Chinese Yuan',    'CNY', 7.2410, 7.8440, '2025-06-15'],
        ['Indian Rupee',    'INR', 83.4500, 90.4000, '2025-06-15'],
        ['Brazilian Real',  'BRL', 4.9870, 5.4030, '2025-06-15'],
        ['Mexican Peso',    'MXN', 17.2300, 18.6600, '2025-06-15'],
    ]

    for r, row_data in enumerate(exchange_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c in [3, 4]:
                cell.number_format = '0.0000'

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 14
    ws2.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Ensure LibreOffice config directory exists but decimal separator is NOT set to comma
    # (default state - period as decimal separator via system locale)
    lo_config_dir = '/home/user/.config/libreoffice/4/user'
    os.makedirs(lo_config_dir, exist_ok=True)

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
