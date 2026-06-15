"""
Reward Script: Calculate population standard deviation in G2
Task ID: calc_fmb_stdev_population_056
Domain: libreoffice_calc
Scoring:
  - Component 1: G2 contains a formula (not empty)            — 0.3 pts
  - Component 2: G2 uses STDEVP (population std dev)          — 0.4 pts
  - Component 3: G2 formula covers exactly D2:D1001 range     — 0.3 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_stdev_population_056'
SHEET_NAME = 'Census Data'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires placing =STDEVP(D2:D1001) in cell G2 of the 'Census Data' sheet.
    Only G2 should change from initial to golden; all other cells remain intact.
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: G2 contains a formula (is not empty) — 0.3 points
    # This FAILS on initial (G2=None) and PASSES on golden (G2='=STDEVP(D2:D1001)')
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str) and g2_value.strip().startswith('='):
            print(f"PASS: Component 1 — G2 contains a formula: {repr(g2_value)} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — G2 should contain a formula, found: {repr(g2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G2 uses STDEVP (population standard deviation, not STDEV/STDEVS) — 0.4 points
    # STDEVP computes population std dev; STDEV/STDEVS compute sample std dev.
    # This FAILS on initial (G2=None) and PASSES on golden (formula contains STDEVP).
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str):
            formula_upper = g2_value.upper().replace(' ', '')
            # Match STDEVP( but NOT STDEVPA( — both are population variants; STDEVP is required
            # Also reject STDEV( and STDEVS( which are sample variants
            if re.search(r'=STDEVP\(', formula_upper) and not re.search(r'=STDEVPA\(', formula_upper):
                print(f"PASS: Component 2 — G2 uses STDEVP (population std dev): {repr(g2_value)} (0.4 pts)")
                total_score += 0.4
            elif re.search(r'=STDEVPA\(', formula_upper):
                print(f"FAIL: Component 2 — G2 uses STDEVPA instead of STDEVP: {repr(g2_value)}")
            elif re.search(r'=STDEV\(', formula_upper) or re.search(r'=STDEVS\(', formula_upper):
                print(f"FAIL: Component 2 — G2 uses sample std dev (STDEV/STDEVS) instead of population STDEVP: {repr(g2_value)}")
            else:
                print(f"FAIL: Component 2 — G2 does not use STDEVP: {repr(g2_value)}")
        else:
            print(f"FAIL: Component 2 — G2 is empty or not a formula: {repr(g2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2 formula covers exactly D2:D1001 (full 1000-row population range) — 0.3 points
    # This FAILS on initial (G2=None) and PASSES on golden (formula contains D2:D1001).
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str):
            formula_upper = g2_value.upper().replace(' ', '')
            # Check for exactly D2:D1001 as the range argument
            if 'D2:D1001' in formula_upper:
                print(f"PASS: Component 3 — G2 formula covers D2:D1001 (all 1000 records): {repr(g2_value)} (0.3 pts)")
                total_score += 0.3
            else:
                # Extract what range is actually used for better feedback
                range_match = re.search(r'STDEVP[A]?\(([^)]+)\)', formula_upper)
                actual_range = range_match.group(1) if range_match else 'unknown'
                print(f"FAIL: Component 3 — Expected range D2:D1001, found: {actual_range} in formula {repr(g2_value)}")
        else:
            print(f"FAIL: Component 3 — G2 is empty or not a formula: {repr(g2_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
