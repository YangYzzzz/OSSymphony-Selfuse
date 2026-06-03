"""
Reward Script: Calculate 401(k) contributions and company match for payroll.
Task ID: calc_fin_401k_contribution_070
Domain: libreoffice_calc
Scoring:
  - Component 1: Employee Contribution formulas D2:D35 (MIN capped at IRS limit) — 0.25 pts
  - Component 2: Company Match formulas E2:E35 (tiered match formula) — 0.25 pts
  - Component 3: Total Benefit formulas F2:F35 (D+E) — 0.15 pts
  - Component 4: Currency formatting on D2:F35 and D36:F36 — 0.10 pts
  - Component 5: Totals row at row 36 (SUM formulas + 'Totals' label + bold) — 0.15 pts
  - Component 6: Row 1 headers bold — 0.05 pts
  - Component 7: Conditional formatting on C2:C35 (orange fill when value=0) — 0.05 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_401k_contribution_070'


def normalize_formula(f):
    """Normalize formula for loose comparison: uppercase, strip spaces."""
    if f is None:
        return ''
    return str(f).upper().replace(' ', '')


def check_min_formula(value):
    """Check if a cell value is a MIN formula that caps with $G$1."""
    if not isinstance(value, str):
        return False
    v = normalize_formula(value)
    # Must start with =MIN( and include $G$1 (IRS limit)
    return v.startswith('=MIN(') and '$G$1' in v


def check_match_formula(value):
    """Check if a cell value is the company match formula (tiered, using MIN/MAX)."""
    if not isinstance(value, str):
        return False
    v = normalize_formula(value)
    # Must start with =MIN( and contain both 0.03 and 0.5
    # Both simplified forms: MIN(B*0.03,B*C)+MIN(B*0.02,MAX(0,...))*.5
    # or IF-based: IF(C>=0.05,B*0.03+B*0.02*0.5, ...)
    return v.startswith('=MIN(') and '0.03' in v and ('0.5' in v or '0.02' in v)


def check_total_formula(value, row):
    """Check if cell is =Dx+Ex formula for given row."""
    if not isinstance(value, str):
        return False
    v = normalize_formula(value)
    expected = f'=D{row}+E{row}'
    return v == expected


def check_currency_format(number_format):
    """Check if number format is a currency format."""
    if not number_format:
        return False
    nf = str(number_format)
    # Accept $#,##0.00 or similar currency patterns
    return '$' in nf and '0.00' in nf


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that '401k' sheet exists — precondition gate
    if '401k' not in wb.sheetnames:
        print("CRITICAL: Sheet '401k' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['401k']

    # Component 1: Employee Contribution formulas D2:D35 (0.25 points)
    # D column: =MIN(Bx*Cx,$G$1) — employee contribution capped at IRS limit ($23,000)
    # This FAILS on initial (D2:D35 are empty) and PASSES on golden
    try:
        d_formula_count = 0
        for row in range(2, 36):
            val = ws.cell(row=row, column=4).value  # column D
            if check_min_formula(val):
                d_formula_count += 1

        if d_formula_count == 34:
            print(f"PASS: Component 1 — All 34 D2:D35 cells have MIN formula capped at IRS limit (0.25 pts)")
            total_score += 0.25
        elif d_formula_count >= 17:
            partial = round(0.25 * (d_formula_count / 34), 4)
            print(f"PARTIAL: Component 1 — {d_formula_count}/34 D-column MIN formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {d_formula_count}/34 D-column MIN formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Company Match formulas E2:E35 (0.25 points)
    # E column: tiered match formula using MIN/MAX — 100% of first 3%, 50% of next 2%
    # This FAILS on initial (E2:E35 are empty) and PASSES on golden
    try:
        e_formula_count = 0
        for row in range(2, 36):
            val = ws.cell(row=row, column=5).value  # column E
            if check_match_formula(val):
                e_formula_count += 1

        if e_formula_count == 34:
            print(f"PASS: Component 2 — All 34 E2:E35 cells have tiered company match formula (0.25 pts)")
            total_score += 0.25
        elif e_formula_count >= 17:
            partial = round(0.25 * (e_formula_count / 34), 4)
            print(f"PARTIAL: Component 2 — {e_formula_count}/34 E-column match formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {e_formula_count}/34 E-column match formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Total Benefit formulas F2:F35 (0.15 points)
    # F column: =Dx+Ex for each row
    # This FAILS on initial (F2:F35 are empty) and PASSES on golden
    try:
        f_formula_count = 0
        for row in range(2, 36):
            val = ws.cell(row=row, column=6).value  # column F
            if check_total_formula(val, row):
                f_formula_count += 1

        if f_formula_count == 34:
            print(f"PASS: Component 3 — All 34 F2:F35 cells have =Dx+Ex total benefit formula (0.15 pts)")
            total_score += 0.15
        elif f_formula_count >= 17:
            partial = round(0.15 * (f_formula_count / 34), 4)
            print(f"PARTIAL: Component 3 — {f_formula_count}/34 F-column total formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {f_formula_count}/34 F-column total formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Currency formatting D2:F35 (0.10 points)
    # Columns D, E, F should be formatted as currency ($#,##0.00)
    # This FAILS on initial (D/E/F cells empty, no currency format) and PASSES on golden
    try:
        currency_count = 0
        total_cells = 0
        for row in range(2, 36):
            for col in range(4, 7):  # columns D, E, F
                total_cells += 1
                cell = ws.cell(row=row, column=col)
                if check_currency_format(cell.number_format):
                    currency_count += 1

        if currency_count == total_cells:
            print(f"PASS: Component 4 — All {total_cells} D2:F35 cells are currency formatted (0.10 pts)")
            total_score += 0.10
        elif currency_count >= total_cells // 2:
            partial = round(0.10 * (currency_count / total_cells), 4)
            print(f"PARTIAL: Component 4 — {currency_count}/{total_cells} currency-formatted cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {currency_count}/{total_cells} cells have currency format")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Totals row at row 36 (0.15 points)
    # A36='Totals' bold, D36=SUM(D2:D35) bold, E36=SUM(E2:E35) bold, F36=SUM(F2:F35) bold
    # This FAILS on initial (row 36 is empty) and PASSES on golden
    try:
        totals_checks = 0

        # Check A36 label
        a36_val = ws.cell(row=36, column=1).value
        a36_bold = ws.cell(row=36, column=1).font.bold
        if a36_val == 'Totals' and a36_bold:
            totals_checks += 1
            print(f"  Sub-check: A36='Totals' bold — PASS")
        else:
            print(f"  Sub-check: A36 expected 'Totals' bold, got {repr(a36_val)}, bold={a36_bold}")

        # Check D36, E36, F36 SUM formulas
        sum_map = {
            4: ('D36', '=SUM(D2:D35)'),
            5: ('E36', '=SUM(E2:E35)'),
            6: ('F36', '=SUM(F2:F35)'),
        }
        for col, (coord, expected_sum) in sum_map.items():
            cell = ws.cell(row=36, column=col)
            cell_val = normalize_formula(cell.value)
            expected_norm = normalize_formula(expected_sum)
            cell_bold = cell.font.bold
            if cell_val == expected_norm and cell_bold:
                totals_checks += 1
                print(f"  Sub-check: {coord}='{expected_sum}' bold — PASS")
            elif cell_val == expected_norm:
                totals_checks += 0  # formula correct but not bold — partial
                print(f"  Sub-check: {coord} formula correct but not bold")
            else:
                print(f"  Sub-check: {coord} expected '{expected_sum}', got {repr(cell.value)}")

        if totals_checks == 4:
            print(f"PASS: Component 5 — Totals row complete with label + 3 SUM formulas (bold) (0.15 pts)")
            total_score += 0.15
        elif totals_checks >= 2:
            partial = round(0.15 * (totals_checks / 4), 4)
            print(f"PARTIAL: Component 5 — {totals_checks}/4 totals sub-checks passed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {totals_checks}/4 totals sub-checks passed")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Row 1 headers bold (0.05 points)
    # Headers in A1:F1 should be bold
    # This FAILS on initial (headers not bold) and PASSES on golden
    try:
        bold_count = 0
        for col in range(1, 7):  # A through F
            cell = ws.cell(row=1, column=col)
            if cell.font.bold:
                bold_count += 1

        if bold_count >= 6:
            print(f"PASS: Component 6 — All A1:F1 headers are bold (0.05 pts)")
            total_score += 0.05
        elif bold_count >= 3:
            partial = round(0.05 * (bold_count / 6), 4)
            print(f"PARTIAL: Component 6 — {bold_count}/6 headers bold ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Only {bold_count}/6 headers are bold")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Conditional formatting on C2:C35 — orange fill when value=0 (0.05 points)
    # This FAILS on initial (no conditional formatting) and PASSES on golden
    try:
        cf_equal_zero_count = 0   # count of cellIs-equal-0 rules found
        cf_orange_count = 0       # count of such rules with orange fill

        for cf_range in ws.conditional_formatting:
            for rule in ws.conditional_formatting[cf_range]:
                # Check if this is a cellIs rule with operator equal and formula 0
                if rule.type == 'cellIs' and rule.operator == 'equal':
                    formula = rule.formula
                    if formula and '0' in str(formula):
                        cf_equal_zero_count += 1
                        # Check if fill color is orange-ish (FFA500)
                        if rule.dxf and rule.dxf.fill:
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                                # FFFFA500 is orange; accept any variant with FFA5
                                if fill_color and 'FFA5' in fill_color.upper():
                                    cf_orange_count += 1
                            except Exception:
                                pass

        if cf_equal_zero_count >= 1 and cf_orange_count >= 1:
            print(f"PASS: Component 7 — Conditional formatting with orange fill for zero contribution found (0.05 pts)")
            total_score += 0.05
        elif cf_equal_zero_count >= 1:
            print(f"PARTIAL: Component 7 — Conditional formatting found but orange color not confirmed (0.025 pts)")
            total_score += 0.025
        else:
            print(f"FAIL: Component 7 — No conditional formatting (cellIs equal 0) found (count={cf_equal_zero_count})")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
