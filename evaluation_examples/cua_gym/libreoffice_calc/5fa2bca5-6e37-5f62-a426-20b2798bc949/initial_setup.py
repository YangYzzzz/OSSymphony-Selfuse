"""
Initial Setup: 401(k) Contribution Calculator Spreadsheet
Task ID: calc_fin_401k_contribution_070
Domain: libreoffice_calc

Creates a 401k payroll spreadsheet with employee data.
Columns A, B, C are filled. D, E, F are empty (to be filled by agent).
G1 contains the IRS annual limit. No formulas, no totals row, no formatting.
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_401k_contribution_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '401k'

    # --- Row 1: Headers (plain, not bold) ---
    headers = ['Employee', 'Annual Salary', 'Contribution %', 'Employee Contribution', 'Company Match', 'Total Benefit']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # G1: IRS annual contribution limit
    ws['G1'] = 23000

    # --- Employee data: 34 employees (rows 2-35) ---
    # Columns A (name), B (annual salary), C (contribution %)
    # D, E, F are intentionally left empty
    employee_data = [
        ('Sarah Chen',        95000,  0.06),
        ('Marcus Johnson',    72000,  0.04),
        ('Emily Rodriguez',   88500,  0.05),
        ('James Whitfield',  110000,  0.10),
        ('Priya Nair',        67000,  0.03),
        ('Derek Okafor',      54000,  0.00),
        ('Linda Kowalski',    82000,  0.07),
        ('Nathan Brooks',     91500,  0.05),
        ('Fatima Al-Hassan', 125000,  0.08),
        ('Tyler Morgan',      61000,  0.02),
        ('Grace Huang',       78000,  0.05),
        ('Samuel Reyes',      69500,  0.04),
        ('Olivia Patel',      97000,  0.06),
        ('Christopher Bell',  58000,  0.00),
        ('Amara Diallo',      83000,  0.03),
        ('Kevin Nakamura',   115000,  0.10),
        ('Rachel Fitzpatrick',76000,  0.05),
        ('Mohammed Hassan',   65000,  0.04),
        ('Stephanie Larson',  92000,  0.07),
        ('David Kim',         87500,  0.06),
        ('Cynthia Torres',    71000,  0.03),
        ('Brian Wallace',     56000,  0.00),
        ('Ngozi Adeyemi',    102000,  0.08),
        ('Patrick Sullivan',  79500,  0.05),
        ('Isabella Rossi',    66000,  0.04),
        ('Aaron Washington',  93000,  0.06),
        ('Mei-Ling Zhou',    118000,  0.10),
        ('Tanya Brennan',     74000,  0.05),
        ('Jerome Mitchell',   59000,  0.00),
        ('Alicia Fernandez',  86000,  0.07),
        ('Ryan Johansson',    70000,  0.04),
        ('Diana Petrov',      81500,  0.05),
        ('Travis Simmons',    63000,  0.03),
        ('Vanessa Okonkwo',   99000,  0.06),
    ]

    for row_idx, (name, salary, contrib_pct) in enumerate(employee_data, 2):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=salary)
        ws.cell(row=row_idx, column=3, value=contrib_pct)
        # D, E, F intentionally left blank

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: 401k')
    print(f'  Rows 2-35: employee data (A, B, C filled; D, E, F empty)')
    print(f'  G1: IRS limit = 23000')
    print(f'  No formulas, no totals row, no bold, no currency format, no conditional formatting')


create_initial()
