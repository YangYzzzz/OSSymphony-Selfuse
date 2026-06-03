"""
Initial Setup: Create event log spreadsheet with oversized column D and undersized column E.
Task ID: calc_gao_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gao_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Events ---
    ws = wb.active
    ws.title = 'Events'

    # Headers
    headers = ['Event ID', 'Event Name', 'Location', 'Event Date', 'Event Description']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - realistic event log entries
    data = [
        ['EVT-001', 'Annual Tech Summit 2025', 'San Francisco, CA', '2025-01-18', 'Three-day conference featuring keynote speakers from leading tech companies and hands-on workshops'],
        ['EVT-002', 'Marketing Strategy Workshop', 'New York, NY', '2025-02-05', 'Half-day intensive session on digital marketing trends and ROI optimization techniques'],
        ['EVT-003', 'Q1 All-Hands Meeting', 'Chicago, IL', '2025-03-12', 'Company-wide quarterly review with department updates and team recognition awards'],
        ['EVT-004', 'Product Launch: CloudSync Pro', 'Austin, TX', '2025-03-28', 'Official launch event for the new enterprise cloud synchronization platform'],
        ['EVT-005', 'HR Benefits Enrollment Fair', 'Denver, CO', '2025-04-10', 'Open enrollment information sessions covering health insurance and retirement plans'],
        ['EVT-006', 'Customer Success Symposium', 'Seattle, WA', '2025-04-22', 'Annual gathering of top-tier clients with product roadmap previews and networking'],
        ['EVT-007', 'Engineering Hackathon', 'Portland, OR', '2025-05-15', 'Weekend coding event focused on AI-powered automation and internal tooling projects'],
        ['EVT-008', 'Sales Kickoff Conference', 'Miami, FL', '2025-06-03', 'Regional sales team alignment meeting with territory planning and quota discussions'],
        ['EVT-009', 'Data Privacy Compliance Seminar', 'Boston, MA', '2025-06-19', 'Mandatory training on GDPR and CCPA regulations for all data-handling personnel'],
        ['EVT-010', 'Summer Team Building Retreat', 'Lake Tahoe, NV', '2025-07-08', 'Three-day offsite retreat with outdoor activities and cross-departmental collaboration exercises'],
        ['EVT-011', 'Board of Directors Meeting', 'New York, NY', '2025-08-14', 'Quarterly board session reviewing financial performance and strategic initiatives'],
        ['EVT-012', 'UX Design Sprint', 'San Diego, CA', '2025-09-02', 'Five-day design sprint to prototype the next-generation mobile application interface'],
        ['EVT-013', 'Cybersecurity Awareness Week', 'Remote', '2025-09-22', 'Organization-wide security training covering phishing, password hygiene, and incident reporting'],
        ['EVT-014', 'Annual Charity Gala', 'Los Angeles, CA', '2025-10-11', 'Black-tie fundraising event supporting STEM education initiatives in underserved communities'],
        ['EVT-015', 'Year-End Planning Summit', 'Chicago, IL', '2025-11-20', 'Leadership offsite for fiscal year planning, budget allocation, and goal setting for 2026'],
    ]

    data_font = Font(name='Calibri', size=11)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 4:  # Date column - center align
                cell.alignment = Alignment(horizontal='center')

    # Set column widths as specified:
    # Column A (Event ID): reasonable width
    ws.column_dimensions['A'].width = 12
    # Column B (Event Name): reasonable width
    ws.column_dimensions['B'].width = 32
    # Column C (Location): reasonable width
    ws.column_dimensions['C'].width = 20
    # Column D (Event Date): TOO WIDE at 5 cm (~18 char units)
    # 1 cm ≈ 3.57 character units in openpyxl
    ws.column_dimensions['D'].width = 18  # ~5 cm, intentionally too wide
    # Column E (Event Description): TOO NARROW at 2 cm (~7 char units)
    ws.column_dimensions['E'].width = 7.14  # ~2 cm, intentionally too narrow

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
