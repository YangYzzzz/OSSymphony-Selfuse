"""
Reward Script: Auto-fit column D and widen column E by 1cm extra
Task ID: calc_gao_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Column D narrowed from initial 18.0 to auto-fit width (~13.0)
  Component 2 (0.4): Column E widened from initial 7.14 to auto-fit + extra (~103.57)
  Component 3 (0.2): Data integrity preserved (all original data intact)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gao_047'

# Known initial column widths (pre-task state)
INITIAL_COL_D_WIDTH = 18.0
INITIAL_COL_E_WIDTH = 7.14


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Events' sheet must exist
    if 'Events' not in wb.sheetnames:
        print("FAIL: 'Events' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Events']

    # Component 1: Column D narrowed (auto-fit) (0.4 points)
    # Initial width was 18.0. After auto-fit for YYYY-MM-DD dates, it should be
    # significantly narrower. Golden value is 13.0. We accept any width that is
    # meaningfully reduced from 18.0 (must be < 16.0) and still reasonable (> 8.0).
    try:
        col_d = ws.column_dimensions.get('D')
        col_d_width = col_d.width if col_d and col_d.width else None

        if col_d_width is not None:
            print(f"INFO: Column D width = {col_d_width} (initial was {INITIAL_COL_D_WIDTH})")
            if col_d_width < 16.0 and col_d_width > 8.0:
                print(f"PASS: Component 1 — Column D narrowed to {col_d_width} (auto-fit range) (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — Column D width {col_d_width} not in expected auto-fit range (8.0, 16.0)")
        else:
            print("FAIL: Component 1 — Column D width not set")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column E significantly widened (auto-fit + 1cm extra) (0.4 points)
    # Initial width was 7.14. The descriptions are very long (60-100+ characters).
    # Auto-fit alone would be very wide. With +1cm it should be substantially wider
    # than 7.14. Golden value is 103.57. We accept width > 50.0 as evidence the
    # column was auto-fitted to content (descriptions are long).
    try:
        col_e = ws.column_dimensions.get('E')
        col_e_width = col_e.width if col_e and col_e.width else None

        if col_e_width is not None:
            print(f"INFO: Column E width = {col_e_width} (initial was {INITIAL_COL_E_WIDTH})")
            if col_e_width > 50.0:
                print(f"PASS: Component 2 — Column E widened to {col_e_width} (auto-fit + extra) (0.4 pts)")
                total_score += 0.4
            elif col_e_width > 20.0:
                # Partial credit: column was widened but maybe not fully auto-fitted
                print(f"PARTIAL: Component 2 — Column E widened to {col_e_width} but less than expected (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 2 — Column E width {col_e_width} not wide enough (need > 50.0)")
        else:
            print("FAIL: Component 2 — Column E width not set")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data integrity (0.2 points)
    # Verify key data cells are still present and unchanged.
    # This is anchored to a change check: we combine it with verifying
    # that column widths actually differ from initial state.
    try:
        # Check that the column widths actually changed from initial values
        col_d_changed = (col_d_width is not None and abs(col_d_width - INITIAL_COL_D_WIDTH) > 1.0)
        col_e_changed = (col_e_width is not None and abs(col_e_width - INITIAL_COL_E_WIDTH) > 1.0)

        if col_d_changed and col_e_changed:
            # Now verify data integrity: spot-check key cells
            header_d = ws['D1'].value
            header_e = ws['E1'].value
            date_val = ws['D2'].value
            desc_val = ws['E2'].value

            failures = []
            if header_d != 'Event Date':
                failures.append(f"D1 expected 'Event Date', got '{header_d}'")
            if header_e != 'Event Description':
                failures.append(f"E1 expected 'Event Description', got '{header_e}'")
            if date_val is None or '2025' not in str(date_val):
                failures.append(f"D2 expected date with '2025', got '{date_val}'")
            if desc_val is None or 'conference' not in str(desc_val).lower():
                failures.append(f"E2 expected description with 'conference', got '{desc_val}'")

            if len(failures) == 0:
                print(f"PASS: Component 3 — Data intact and column widths changed from initial (0.2 pts)")
                total_score += 0.2
            else:
                for f in failures:
                    print(f"FAIL: Component 3 — {f}")
        else:
            print(f"FAIL: Component 3 — Column widths did not change from initial (D changed: {col_d_changed}, E changed: {col_e_changed})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
