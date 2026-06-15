"""
Initial Setup: VLOOKUP department names from lookup table
Task ID: calc_gg5_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Departments sheet (lookup table) ---
    ws_dept = wb.active
    ws_dept.title = 'Departments'
    dept_headers = ['Dept ID', 'Dept Name']
    for c, h in enumerate(dept_headers, 1):
        ws_dept.cell(row=1, column=c, value=h)

    departments = [
        (101, 'Engineering'),
        (102, 'Marketing'),
        (103, 'Finance'),
        (104, 'Human Resources'),
        (105, 'Sales'),
        (106, 'Operations'),
        (107, 'Legal'),
        (108, 'Product'),
        (109, 'Customer Support'),
        (110, 'Research'),
        (111, 'Design'),
        (112, 'IT Infrastructure'),
    ]
    for r, (did, dname) in enumerate(departments, 2):
        ws_dept.cell(row=r, column=1, value=did)
        ws_dept.cell(row=r, column=2, value=dname)

    ws_dept.column_dimensions['A'].width = 12
    ws_dept.column_dimensions['B'].width = 22

    # --- Employees sheet ---
    ws_emp = wb.create_sheet('Employees', 0)  # insert first

    # Headers: A=Employee ID, B=Name, C=Title, D=Location, E=Salary, F=Join Date, G=Manager ID
    # H is intentionally left empty (task asks agent to add VLOOKUP and "Department" header)
    emp_headers = ['Employee ID', 'Name', 'Title', 'Location', 'Salary', 'Join Date', 'Manager ID']
    for c, h in enumerate(emp_headers, 1):
        ws_emp.cell(row=1, column=c, value=h)

    first_names = [
        'Sarah', 'Marcus', 'Priya', 'James', 'Yuki', 'Carlos', 'Amina', 'David',
        'Mei', 'Robert', 'Fatima', 'Alexander', 'Lucia', 'Benjamin', 'Aisha',
        'Thomas', 'Sofia', 'Omar', 'Elena', 'Nathan', 'Grace', 'Ibrahim',
        'Hannah', 'Wei', 'Rachel', 'Daniel', 'Olivia', 'Raj', 'Emma', 'Liam',
        'Zara', 'Ethan', 'Nora', 'Kai', 'Isla', 'Mason', 'Chloe', 'Arjun',
        'Mia', 'Leo', 'Ava', 'Ravi', 'Lily', 'Owen', 'Jasmine', 'Caleb',
        'Layla', 'Dylan', 'Mila', 'Finn',
    ]
    last_names = [
        'Chen', 'Johnson', 'Patel', 'Williams', 'Tanaka', 'Rodriguez', 'Hassan',
        'Kim', 'Zhang', 'Smith', 'Ali', 'Petrov', 'Garcia', 'Brown', 'Ibrahim',
        'Anderson', 'Mendez', 'Martinez', 'Kowalski', 'Nakamura', 'Taylor',
        'Singh', 'Lee', 'Wang', 'Davis', 'Moore', 'Jackson', 'White', 'Lopez',
        'Thompson', 'Okafor', 'Sato', 'Novak', 'Fernandez', 'Schmidt',
    ]
    titles = [
        'Software Engineer', 'Senior Software Engineer', 'Staff Engineer',
        'Marketing Manager', 'Financial Analyst', 'Senior Financial Analyst',
        'HR Specialist', 'Sales Representative', 'Senior Sales Rep',
        'Operations Analyst', 'Legal Counsel', 'Product Manager',
        'Support Specialist', 'Senior Support Lead', 'Research Scientist',
        'UX Designer', 'Senior UX Designer', 'IT Administrator',
        'Data Analyst', 'Project Manager', 'Business Analyst',
        'Account Executive', 'QA Engineer', 'DevOps Engineer',
        'Technical Writer',
    ]
    locations = [
        'New York', 'San Francisco', 'Chicago', 'Austin', 'Seattle',
        'Boston', 'Denver', 'Atlanta', 'Portland', 'Miami',
        'Los Angeles', 'Dallas', 'Phoenix', 'Minneapolis', 'Detroit',
    ]
    dept_ids = [d[0] for d in departments]

    base_date = date(2018, 1, 1)
    date_range = (date(2025, 6, 1) - base_date).days

    for i in range(149):
        row = i + 2
        # Employee ID is the dept ID for this employee (so VLOOKUP(A2, Departments.$A:$B, 2, FALSE) works)
        # Each employee is assigned to one of the 12 departments
        emp_dept_id = dept_ids[i % len(dept_ids)]

        fname = first_names[i % len(first_names)]
        lname = last_names[(i * 3) % len(last_names)]
        name = f'{fname} {lname}'
        title = titles[i % len(titles)]
        location = locations[i % len(locations)]
        salary = round(random.randint(48000, 165000), -2)
        join_date = base_date + timedelta(days=random.randint(0, date_range))
        manager_id = dept_ids[random.randint(0, 11)] if i > 0 else ''

        ws_emp.cell(row=row, column=1, value=emp_dept_id)   # A: Employee ID (actually dept ID for VLOOKUP)
        ws_emp.cell(row=row, column=2, value=name)           # B: Name
        ws_emp.cell(row=row, column=3, value=title)          # C: Title
        ws_emp.cell(row=row, column=4, value=location)       # D: Location
        ws_emp.cell(row=row, column=5, value=salary)         # E: Salary
        ws_emp.cell(row=row, column=6, value=join_date)      # F: Join Date
        if manager_id != '':
            ws_emp.cell(row=row, column=7, value=manager_id) # G: Manager ID
        # H: intentionally empty — task requires VLOOKUP formulas here

    # Column widths
    ws_emp.column_dimensions['A'].width = 14
    ws_emp.column_dimensions['B'].width = 22
    ws_emp.column_dimensions['C'].width = 26
    ws_emp.column_dimensions['D'].width = 16
    ws_emp.column_dimensions['E'].width = 12
    ws_emp.column_dimensions['F'].width = 14
    ws_emp.column_dimensions['G'].width = 14
    ws_emp.column_dimensions['H'].width = 20

    # Format salary column
    for r in range(2, 151):
        ws_emp.cell(row=r, column=5).number_format = '$#,##0'

    # Format date column
    for r in range(2, 151):
        ws_emp.cell(row=r, column=6).number_format = 'yyyy-mm-dd'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
