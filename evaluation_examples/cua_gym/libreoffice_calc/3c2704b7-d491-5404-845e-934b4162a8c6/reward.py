"""
Reward Script: VLOOKUP department names in hr_data.xlsx
Task ID: calc_gg5_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): H1 header is 'Department'
  Component 2 (0.45): H2:H150 all contain VLOOKUP formulas
  Component 3 (0.20): VLOOKUP formulas reference Departments sheet with absolute refs
  Component 4 (0.20): VLOOKUP uses exact match (FALSE or 0 as 4th arg)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_012'


def persist_app_state(domain: str):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Employees sheet exists
    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: 'Employees' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Employees']

    # Component 1: H1 header is 'Department' (0.15 points)
    try:
        h1_val = ws['H1'].value
        if h1_val is not None and str(h1_val).strip().lower() == 'department':
            print(f"PASS: Component 1 — H1 header is '{h1_val}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — expected 'Department' in H1, found: {repr(h1_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: H2:H150 contain VLOOKUP formulas (0.45 points)
    # Check that a high proportion of cells have VLOOKUP formulas
    try:
        vlookup_count = 0
        total_cells = 149  # H2 through H150
        for row in range(2, 151):
            val = ws.cell(row=row, column=8).value
            if val is not None and isinstance(val, str) and 'VLOOKUP' in val.upper():
                vlookup_count += 1

        ratio = vlookup_count / total_cells
        if ratio >= 0.95:
            print(f"PASS: Component 2 — {vlookup_count}/{total_cells} cells have VLOOKUP formulas (0.45 pts)")
            total_score += 0.45
        elif ratio >= 0.5:
            partial = 0.45 * (ratio - 0.5) / 0.5
            print(f"PARTIAL: Component 2 — {vlookup_count}/{total_cells} VLOOKUP formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — only {vlookup_count}/{total_cells} cells have VLOOKUP formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: VLOOKUP references Departments sheet with absolute refs (0.20 points)
    # Check a sample of formulas for correct reference pattern
    try:
        correct_ref_count = 0
        sample_rows = [2, 10, 50, 100, 150]
        for row in sample_rows:
            val = ws.cell(row=row, column=8).value
            if val is not None and isinstance(val, str):
                upper_val = val.upper().replace(' ', '')
                # Check for Departments sheet reference with $ (absolute refs)
                # Acceptable patterns: Departments.$A:$B, Departments!$A:$B, etc.
                if 'DEPARTMENTS' in upper_val and '$' in upper_val:
                    correct_ref_count += 1

        if correct_ref_count >= 4:
            print(f"PASS: Component 3 — {correct_ref_count}/{len(sample_rows)} sampled formulas have absolute Departments refs (0.20 pts)")
            total_score += 0.20
        elif correct_ref_count >= 2:
            partial = 0.20 * correct_ref_count / len(sample_rows)
            print(f"PARTIAL: Component 3 — {correct_ref_count}/{len(sample_rows)} sampled formulas correct ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — only {correct_ref_count}/{len(sample_rows)} sampled formulas have absolute Departments refs")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: VLOOKUP uses exact match — FALSE or 0 as 4th argument (0.20 points)
    try:
        exact_match_count = 0
        sample_rows = [2, 10, 50, 100, 150]
        for row in sample_rows:
            val = ws.cell(row=row, column=8).value
            if val is not None and isinstance(val, str):
                upper_val = val.upper().replace(' ', '')
                # The 4th argument should be FALSE or 0 for exact match
                # Pattern: VLOOKUP(..., FALSE) or VLOOKUP(..., 0)
                if re.search(r'VLOOKUP\(.*,(FALSE|0)\)', upper_val):
                    exact_match_count += 1

        if exact_match_count >= 4:
            print(f"PASS: Component 4 — {exact_match_count}/{len(sample_rows)} sampled formulas use exact match (0.20 pts)")
            total_score += 0.20
        elif exact_match_count >= 2:
            partial = 0.20 * exact_match_count / len(sample_rows)
            print(f"PARTIAL: Component 4 — {exact_match_count}/{len(sample_rows)} sampled formulas use exact match ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — only {exact_match_count}/{len(sample_rows)} sampled formulas use exact match")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
