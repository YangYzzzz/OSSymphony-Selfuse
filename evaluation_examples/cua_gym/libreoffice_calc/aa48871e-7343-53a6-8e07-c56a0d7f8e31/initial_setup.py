"""
Initial Setup: Warehouse picking list generator
Task ID: calc_wf_025
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Master Inventory (50 items) ---
    ws1 = wb.active
    ws1.title = 'Master Inventory'

    headers1 = ['Location', 'SKU', 'Product', 'Qty Available']
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)

    # 50 inventory items across 5 zones
    inventory_data = [
        ('Z01-R01-S01', 'SKU-1001', 'Industrial Bolt M8x40', 2500),
        ('Z01-R01-S02', 'SKU-1002', 'Hex Nut M8 Zinc', 5000),
        ('Z01-R01-S03', 'SKU-1003', 'Flat Washer M8 SS', 8000),
        ('Z01-R02-S01', 'SKU-1004', 'Spring Washer M10', 3200),
        ('Z01-R02-S02', 'SKU-1005', 'Socket Head Cap Screw M6x25', 4100),
        ('Z01-R03-S01', 'SKU-1006', 'Threaded Rod M12x1000', 150),
        ('Z01-R03-S02', 'SKU-1007', 'Wing Nut M6 Brass', 6200),
        ('Z01-R04-S01', 'SKU-1008', 'Lock Nut M10 Nylon Insert', 3800),
        ('Z01-R04-S02', 'SKU-1009', 'Carriage Bolt M8x50', 1900),
        ('Z01-R05-S01', 'SKU-1010', 'Eye Bolt M10x80', 750),
        ('Z02-R01-S01', 'SKU-2001', 'PVC Pipe 2in x 10ft', 320),
        ('Z02-R01-S02', 'SKU-2002', 'Copper Elbow 3/4in 90deg', 1800),
        ('Z02-R01-S03', 'SKU-2003', 'Ball Valve 1in Brass', 420),
        ('Z02-R02-S01', 'SKU-2004', 'Pipe Clamp 2in Galvanized', 2600),
        ('Z02-R02-S02', 'SKU-2005', 'Teflon Tape 1/2in x 520in', 5500),
        ('Z02-R03-S01', 'SKU-2006', 'Pipe Reducer 2in to 1in', 890),
        ('Z02-R03-S02', 'SKU-2007', 'Check Valve 1in PVC', 340),
        ('Z02-R04-S01', 'SKU-2008', 'Flexible Hose 3/4in x 6ft', 280),
        ('Z02-R04-S02', 'SKU-2009', 'Pipe Thread Sealant 8oz', 1200),
        ('Z02-R05-S01', 'SKU-2010', 'Gate Valve 1.5in Cast Iron', 195),
        ('Z03-R01-S01', 'SKU-3001', 'LED Panel Light 2x4ft 40W', 160),
        ('Z03-R01-S02', 'SKU-3002', 'Wire Nut Yellow 100pk', 4800),
        ('Z03-R01-S03', 'SKU-3003', 'Romex 14/2 NM-B 250ft', 85),
        ('Z03-R02-S01', 'SKU-3004', 'Single Pole Switch 15A', 2200),
        ('Z03-R02-S02', 'SKU-3005', 'Duplex Receptacle 15A', 3100),
        ('Z03-R03-S01', 'SKU-3006', 'Circuit Breaker 20A SP', 480),
        ('Z03-R03-S02', 'SKU-3007', 'Conduit EMT 3/4in x 10ft', 650),
        ('Z03-R04-S01', 'SKU-3008', 'Junction Box 4x4 Metal', 1750),
        ('Z03-R04-S02', 'SKU-3009', 'Cable Staple 1/2in 100pk', 6800),
        ('Z03-R05-S01', 'SKU-3010', 'Outdoor Flood Light 50W', 220),
        ('Z04-R01-S01', 'SKU-4001', 'Drywall Sheet 4x8ft 1/2in', 240),
        ('Z04-R01-S02', 'SKU-4002', 'Joint Compound 5gal', 180),
        ('Z04-R01-S03', 'SKU-4003', 'Drywall Tape 250ft Roll', 950),
        ('Z04-R02-S01', 'SKU-4004', 'Wood Stud 2x4x8ft SPF', 1800),
        ('Z04-R02-S02', 'SKU-4005', 'Plywood Sheet 4x8ft 3/4in', 130),
        ('Z04-R03-S01', 'SKU-4006', 'OSB Sheathing 4x8ft 7/16in', 200),
        ('Z04-R03-S02', 'SKU-4007', 'Insulation Batt R-13 15in', 320),
        ('Z04-R04-S01', 'SKU-4008', 'Construction Adhesive 10oz', 2800),
        ('Z04-R04-S02', 'SKU-4009', 'Framing Nail 3in 5lb Box', 1600),
        ('Z04-R05-S01', 'SKU-4010', 'Deck Screw #8x2.5in 1lb', 3400),
        ('Z05-R01-S01', 'SKU-5001', 'Exterior Latex Paint 1gal White', 380),
        ('Z05-R01-S02', 'SKU-5002', 'Interior Semi-Gloss 1gal Beige', 450),
        ('Z05-R01-S03', 'SKU-5003', 'Paint Roller 9in Microfiber', 1200),
        ('Z05-R02-S01', 'SKU-5004', 'Painter Tape 1.5in x 60yd Blue', 2800),
        ('Z05-R02-S02', 'SKU-5005', 'Drop Cloth 9x12ft Canvas', 340),
        ('Z05-R03-S01', 'SKU-5006', 'Sandpaper 220 Grit 9x11 25pk', 1500),
        ('Z05-R03-S02', 'SKU-5007', 'Wood Stain 1qt Dark Walnut', 280),
        ('Z05-R04-S01', 'SKU-5008', 'Polyurethane 1qt Satin', 320),
        ('Z05-R04-S02', 'SKU-5009', 'Caulk Silicone 10.1oz White', 4200),
        ('Z05-R05-S01', 'SKU-5010', 'Spray Paint 12oz Gloss Black', 1800),
    ]

    for r, row_data in enumerate(inventory_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet 2: Orders (15 items to pick) ---
    ws2 = wb.create_sheet('Orders')
    headers2 = ['Order #', 'SKU', 'Qty Ordered']
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)

    # 15 order items spanning multiple zones
    order_data = [
        ('ORD-4521', 'SKU-1001', 50),
        ('ORD-4521', 'SKU-2003', 8),
        ('ORD-4521', 'SKU-3005', 24),
        ('ORD-4522', 'SKU-5001', 5),
        ('ORD-4522', 'SKU-1007', 100),
        ('ORD-4522', 'SKU-4004', 30),
        ('ORD-4523', 'SKU-2007', 12),
        ('ORD-4523', 'SKU-3001', 6),
        ('ORD-4523', 'SKU-4009', 10),
        ('ORD-4524', 'SKU-1003', 200),
        ('ORD-4524', 'SKU-2005', 50),
        ('ORD-4524', 'SKU-5006', 15),
        ('ORD-4525', 'SKU-3008', 20),
        ('ORD-4525', 'SKU-4001', 8),
        ('ORD-4525', 'SKU-5009', 36),
    ]

    for r, row_data in enumerate(order_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Sheet 3: Pick List (empty - agent needs to build this) ---
    ws3 = wb.create_sheet('Pick List')
    # Only sheet exists, no headers, no formulas, no formatting
    # The task requires the agent to populate this sheet

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
