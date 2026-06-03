"""
Reward Script: Warehouse Picking List Generator
Task ID: calc_wf_025
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Pick List headers and structure (6 columns, 15 data rows + total row)
  Component 2 (0.25): Lookup formulas (VLOOKUP/INDEX-MATCH) for Location, Product, Available Stock
  Component 3 (0.20): Data completeness - all 15 order SKUs present with correct quantities
  Component 4 (0.10): Items sorted by zone (SKU prefix ordering)
  Component 5 (0.10): Formatting - bold headers, header fill color, freeze panes
  Component 6 (0.10): Print area set and checkbox column present
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_025'

# Expected SKUs from Orders sheet, sorted by zone
EXPECTED_SKUS = [
    'SKU-1001', 'SKU-1003', 'SKU-1007',
    'SKU-2003', 'SKU-2005', 'SKU-2007',
    'SKU-3001', 'SKU-3005', 'SKU-3008',
    'SKU-4001', 'SKU-4004', 'SKU-4009',
    'SKU-5001', 'SKU-5006', 'SKU-5009',
]

# Expected quantities from Orders
EXPECTED_QTY = {
    'SKU-1001': 50, 'SKU-1003': 200, 'SKU-1007': 100,
    'SKU-2003': 8, 'SKU-2005': 50, 'SKU-2007': 12,
    'SKU-3001': 6, 'SKU-3005': 24, 'SKU-3008': 20,
    'SKU-4001': 8, 'SKU-4004': 30, 'SKU-4009': 10,
    'SKU-5001': 5, 'SKU-5006': 15, 'SKU-5009': 36,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Pick List' sheet exists
    if 'Pick List' not in wb.sheetnames:
        print("FAIL: 'Pick List' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pick List']

    # =========================================================================
    # Component 1: Pick List headers and structure (0.25 points)
    # In initial_env the Pick List sheet is empty (A1=None, max_row=1).
    # Golden has headers in row 1, 15 data rows (2-16), and a total row (17).
    # =========================================================================
    try:
        comp1 = 0.0
        # Check headers exist in row 1 (at least 5 columns with text)
        headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
        non_empty_headers = [h for h in headers if h is not None]
        if len(non_empty_headers) >= 5:
            # Check that expected header-like names are present
            header_lower = [str(h).lower() for h in non_empty_headers]
            has_location = any('location' in h for h in header_lower)
            has_sku = any('sku' in h for h in header_lower)
            has_product = any('product' in h for h in header_lower)
            has_qty = any('qty' in h or 'pick' in h or 'quantity' in h for h in header_lower)
            if has_location and has_sku and has_product and has_qty:
                comp1 += 0.10
                print(f"PASS: Headers contain Location, SKU, Product, Qty ({headers})")
            else:
                print(f"FAIL: Missing expected headers. Found: {headers}")
        else:
            print(f"FAIL: Not enough headers. Found: {headers}")

        # Check data rows: at least 10 rows with data below header
        data_row_count = 0
        for r in range(2, ws.max_row + 1):
            sku_val = ws.cell(row=r, column=2).value  # SKU column
            if sku_val is not None and str(sku_val).startswith('SKU-'):
                data_row_count += 1
        if data_row_count >= 15:
            comp1 += 0.10
            print(f"PASS: Found {data_row_count} data rows with SKUs (expected 15)")
        elif data_row_count >= 10:
            comp1 += 0.05
            print(f"PARTIAL: Found {data_row_count} data rows (expected 15)")
        else:
            print(f"FAIL: Only {data_row_count} data rows found (expected 15)")

        # Check total row exists (SUM formula in column D somewhere after data)
        total_found = 0  # 0=not found, 1=found
        for r in range(data_row_count + 2, ws.max_row + 1):
            d_val = ws.cell(row=r, column=4).value
            c_val = ws.cell(row=r, column=3).value
            if d_val is not None and isinstance(d_val, str) and 'SUM' in d_val.upper():
                total_found = 1
                break
            if c_val is not None and 'total' in str(c_val).lower():
                total_found = 1
                break
        if total_found > 0:
            comp1 += 0.05
            print(f"PASS: Total row found")
        else:
            print(f"FAIL: No total/SUM row found")

        if comp1 > 0:
            total_score += comp1
        print(f"Component 1 total: {comp1}/0.25")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Lookup formulas for Location, Product, Available Stock (0.25 points)
    # Initial Pick List has no formulas. Golden uses INDEX-MATCH and VLOOKUP.
    # =========================================================================
    try:
        comp2 = 0.0
        formula_checks = {'location': 0, 'product': 0, 'stock': 0}
        sample_rows = range(2, min(17, ws.max_row + 1))

        for r in sample_rows:
            # Column A (Location): should have INDEX-MATCH or VLOOKUP formula
            loc_val = ws.cell(row=r, column=1).value
            if loc_val is not None and isinstance(loc_val, str):
                loc_upper = loc_val.upper()
                if 'INDEX' in loc_upper or 'VLOOKUP' in loc_upper or 'MATCH' in loc_upper:
                    formula_checks['location'] += 1

            # Column C (Product): should have VLOOKUP or INDEX-MATCH formula
            prod_val = ws.cell(row=r, column=3).value
            if prod_val is not None and isinstance(prod_val, str):
                prod_upper = prod_val.upper()
                if 'VLOOKUP' in prod_upper or 'INDEX' in prod_upper:
                    formula_checks['product'] += 1

            # Column E (Available Stock): should have VLOOKUP formula
            stock_val = ws.cell(row=r, column=5).value
            if stock_val is not None and isinstance(stock_val, str):
                stock_upper = stock_val.upper()
                if 'VLOOKUP' in stock_upper or 'INDEX' in stock_upper:
                    formula_checks['stock'] += 1

        total_data_rows = len(list(sample_rows))
        threshold = max(1, total_data_rows * 0.7)  # 70% of rows should have formulas

        if formula_checks['location'] >= threshold:
            comp2 += 0.10
            print(f"PASS: Location lookup formulas found in {formula_checks['location']}/{total_data_rows} rows")
        else:
            print(f"FAIL: Location lookup formulas in only {formula_checks['location']}/{total_data_rows} rows")

        if formula_checks['product'] >= threshold:
            comp2 += 0.08
            print(f"PASS: Product lookup formulas found in {formula_checks['product']}/{total_data_rows} rows")
        else:
            print(f"FAIL: Product lookup formulas in only {formula_checks['product']}/{total_data_rows} rows")

        if formula_checks['stock'] >= threshold:
            comp2 += 0.07
            print(f"PASS: Available Stock lookup formulas found in {formula_checks['stock']}/{total_data_rows} rows")
        else:
            print(f"FAIL: Stock lookup formulas in only {formula_checks['stock']}/{total_data_rows} rows")

        if comp2 > 0:
            total_score += comp2
        print(f"Component 2 total: {comp2}/0.25")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Data completeness - all 15 SKUs with correct quantities (0.20 points)
    # Initial Pick List has no data. Golden has all 15 order SKUs.
    # =========================================================================
    try:
        comp3 = 0.0
        found_skus = {}
        for r in range(2, ws.max_row + 1):
            sku = ws.cell(row=r, column=2).value
            qty = ws.cell(row=r, column=4).value
            if sku is not None and str(sku).startswith('SKU-'):
                found_skus[str(sku)] = qty

        # Check SKU coverage
        matched_skus = set(found_skus.keys()) & set(EXPECTED_SKUS)
        sku_ratio = len(matched_skus) / len(EXPECTED_SKUS)
        if sku_ratio >= 1.0:
            comp3 += 0.10
            print(f"PASS: All 15 expected SKUs found")
        elif sku_ratio >= 0.7:
            comp3 += 0.05
            print(f"PARTIAL: {len(matched_skus)}/15 expected SKUs found")
        else:
            print(f"FAIL: Only {len(matched_skus)}/15 expected SKUs found")

        # Check quantities match
        qty_matches = 0
        for sku in matched_skus:
            expected_q = EXPECTED_QTY.get(sku)
            actual_q = found_skus.get(sku)
            if expected_q is not None and actual_q is not None:
                try:
                    if abs(float(actual_q) - float(expected_q)) < 0.01:
                        qty_matches += 1
                except (ValueError, TypeError):
                    pass
        if len(matched_skus) > 0:
            qty_ratio = qty_matches / len(EXPECTED_SKUS)
            if qty_ratio >= 0.9:
                comp3 += 0.10
                print(f"PASS: {qty_matches}/15 quantities match")
            elif qty_ratio >= 0.5:
                comp3 += 0.05
                print(f"PARTIAL: {qty_matches}/15 quantities match")
            else:
                print(f"FAIL: Only {qty_matches}/15 quantities match")
        else:
            print(f"FAIL: No matched SKUs to check quantities")

        if comp3 > 0:
            total_score += comp3
        print(f"Component 3 total: {comp3}/0.20")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =========================================================================
    # Component 4: Items sorted by zone (0.10 points)
    # In initial Pick List there's no data, so no sorting. Golden sorts by SKU zone.
    # =========================================================================
    try:
        comp4 = 0.0
        skus_in_order = []
        for r in range(2, ws.max_row + 1):
            sku = ws.cell(row=r, column=2).value
            if sku is not None and str(sku).startswith('SKU-'):
                skus_in_order.append(str(sku))

        if len(skus_in_order) >= 10:
            # Extract zone numbers from SKUs (SKU-1xxx -> zone 1, SKU-2xxx -> zone 2, etc.)
            zones = []
            for sku in skus_in_order:
                try:
                    zone_digit = sku.split('-')[1][0]  # first digit after SKU-
                    zones.append(int(zone_digit))
                except (IndexError, ValueError):
                    zones.append(999)

            # Check if zones are in non-decreasing order
            is_sorted = all(zones[i] <= zones[i+1] for i in range(len(zones)-1))
            if is_sorted:
                comp4 = 0.10
                print(f"PASS: Items sorted by zone order: {zones}")
            else:
                print(f"FAIL: Items not sorted by zone. Zone order: {zones}")
        else:
            print(f"FAIL: Not enough data rows to check sorting ({len(skus_in_order)} rows)")

        if comp4 > 0:
            total_score += comp4
        print(f"Component 4 total: {comp4}/0.10")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =========================================================================
    # Component 5: Formatting - bold headers, header fill, freeze panes (0.10 points)
    # Initial Pick List has no formatting. Golden has bold headers, blue fill, freeze at A2.
    # =========================================================================
    try:
        comp5 = 0.0

        # Check bold headers
        bold_count = 0
        for c in range(1, 7):
            cell = ws.cell(row=1, column=c)
            if cell.value is not None and cell.font and cell.font.bold:
                bold_count += 1
        if bold_count >= 4:
            comp5 += 0.03
            print(f"PASS: {bold_count} bold headers found")
        else:
            print(f"FAIL: Only {bold_count} bold headers (expected >=4)")

        # Check header fill color (any non-default fill)
        fill_count = 0
        for c in range(1, 7):
            cell = ws.cell(row=1, column=c)
            if cell.value is not None and cell.fill and cell.fill.fill_type == 'solid':
                fill_count += 1
        if fill_count >= 4:
            comp5 += 0.03
            print(f"PASS: {fill_count} headers have fill color")
        else:
            print(f"FAIL: Only {fill_count} headers have fill color (expected >=4)")

        # Check freeze panes
        if ws.freeze_panes is not None:
            comp5 += 0.04
            print(f"PASS: Freeze panes set at {ws.freeze_panes}")
        else:
            print(f"FAIL: No freeze panes set")

        if comp5 > 0:
            total_score += comp5
        print(f"Component 5 total: {comp5}/0.10")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # =========================================================================
    # Component 6: Print area set and checkbox column present (0.10 points)
    # Initial Pick List has no print area and no checkbox column.
    # =========================================================================
    try:
        comp6 = 0.0

        # Check print area is set
        if ws.print_area and len(str(ws.print_area)) > 0:
            comp6 += 0.05
            print(f"PASS: Print area set to: {ws.print_area}")
        else:
            print(f"FAIL: No print area set")

        # Check checkbox column (column F "Picked" or similar - should be empty cells for manual use)
        # The column header should indicate it's for checking/picking
        checkbox_col_found = 0  # 0=not found, 1=found
        headers_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        for h in headers_row:
            if h is not None:
                h_lower = str(h).lower()
                if any(kw in h_lower for kw in ['picked', 'check', 'done', 'complete', 'mark', 'tick']):
                    checkbox_col_found = 1
                    break
        if checkbox_col_found > 0:
            comp6 += 0.05
            print(f"PASS: Checkbox/Picked column found in headers: {headers_row}")
        else:
            print(f"FAIL: No checkbox/picked column found. Headers: {headers_row}")

        if comp6 > 0:
            total_score += comp6
        print(f"Component 6 total: {comp6}/0.10")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
