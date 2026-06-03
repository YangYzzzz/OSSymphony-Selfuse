"""
Initial Setup: missing_dois.ods - Calc spreadsheet with 6 ML paper titles and empty DOI column
Task ID: osworld_multi_apps_web_references_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_references_005'
DESKTOP = f'{WORKDIR}/Desktop'
OUTPUT = f'{DESKTOP}/missing_dois.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    # Use Python script to create ODS via odfpy (or fallback to xlsx converted by LibreOffice)
    # First try odfpy
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties, TableColumnProperties
        import odf.namespaces

        doc = OpenDocumentSpreadsheet()

        # Create a sheet named "Sheet1"
        sheet = Table(name="Sheet1")

        def make_cell(value, value_type="string"):
            tc = TableCell()
            if value_type == "string":
                tc.setAttribute("valuetype", "string")
            elif value_type == "float":
                tc.setAttribute("valuetype", "float")
                tc.setAttribute("value", str(value))
            p = P(text=str(value) if value is not None else "")
            tc.addElement(p)
            return tc

        def make_empty_cell():
            tc = TableCell()
            return tc

        # Header row
        header_row = TableRow()
        for h in ["Title", "Authors", "Year", "Venue", "DOI"]:
            header_row.addElement(make_cell(h))
        sheet.addElement(header_row)

        # Data rows - 6 ML papers with empty DOI column
        papers = [
            (
                "Playing Atari with Deep Reinforcement Learning",
                "Mnih et al.",
                2013,
                "arXiv preprint"
            ),
            (
                "Generative Adversarial Networks",
                "Goodfellow et al.",
                2014,
                "NeurIPS"
            ),
            (
                "Dropout: A Simple Way to Prevent Neural Networks from Overfitting",
                "Srivastava et al.",
                2014,
                "Journal of Machine Learning Research"
            ),
            (
                "Adam: A Method for Stochastic Optimization",
                "Kingma, Ba",
                2015,
                "ICLR"
            ),
            (
                "Batch Normalization: Accelerating Deep Network Training by Reducing Internal Covariate Shift",
                "Ioffe, Szegedy",
                2015,
                "ICML"
            ),
            (
                "ImageNet Classification with Deep Convolutional Neural Networks",
                "Krizhevsky et al.",
                2012,
                "NeurIPS"
            ),
        ]

        for title, authors, year, venue in papers:
            row = TableRow()
            row.addElement(make_cell(title))
            row.addElement(make_cell(authors))
            # Year as float cell
            year_cell = TableCell()
            year_cell.setAttribute("valuetype", "float")
            year_cell.setAttribute("value", str(year))
            year_cell.addElement(P(text=str(year)))
            row.addElement(year_cell)
            row.addElement(make_cell(venue))
            row.addElement(make_empty_cell())  # DOI - empty
            sheet.addElement(row)

        doc.spreadsheet.addElement(sheet)
        doc.save(OUTPUT)
        print(f"Initial ODS file created via odfpy: {OUTPUT}")

    except ImportError:
        # Fallback: create xlsx with openpyxl, then convert to ods via LibreOffice
        import openpyxl

        tmp_xlsx = f'{DESKTOP}/missing_dois_tmp.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Headers
        headers = ["Title", "Authors", "Year", "Venue", "DOI"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        # Data rows with empty DOI column
        papers = [
            (
                "Playing Atari with Deep Reinforcement Learning",
                "Mnih et al.",
                2013,
                "arXiv preprint",
                None
            ),
            (
                "Generative Adversarial Networks",
                "Goodfellow et al.",
                2014,
                "NeurIPS",
                None
            ),
            (
                "Dropout: A Simple Way to Prevent Neural Networks from Overfitting",
                "Srivastava et al.",
                2014,
                "Journal of Machine Learning Research",
                None
            ),
            (
                "Adam: A Method for Stochastic Optimization",
                "Kingma, Ba",
                2015,
                "ICLR",
                None
            ),
            (
                "Batch Normalization: Accelerating Deep Network Training by Reducing Internal Covariate Shift",
                "Ioffe, Szegedy",
                2015,
                "ICML",
                None
            ),
            (
                "ImageNet Classification with Deep Convolutional Neural Networks",
                "Krizhevsky et al.",
                2012,
                "NeurIPS",
                None
            ),
        ]

        for r, row_data in enumerate(papers, 2):
            for c, val in enumerate(row_data, 1):
                if val is not None:
                    ws.cell(row=r, column=c, value=val)

        wb.save(tmp_xlsx)
        print(f"Temporary XLSX created: {tmp_xlsx}")

        # Convert to ODS via LibreOffice headless
        env = os.environ.copy()
        env["DISPLAY"] = ":0"
        result = subprocess.run(
            [
                "libreoffice",
                "--headless",
                "--convert-to", "ods",
                "--outdir", DESKTOP,
                tmp_xlsx
            ],
            capture_output=True,
            text=True,
            env=env,
            timeout=60
        )
        print(f"LibreOffice convert stdout: {result.stdout}")
        print(f"LibreOffice convert stderr: {result.stderr}")

        # Rename converted file to correct name
        converted = f'{DESKTOP}/missing_dois_tmp.ods'
        if os.path.exists(converted):
            os.rename(converted, OUTPUT)
            print(f"Renamed to: {OUTPUT}")
        elif os.path.exists(OUTPUT):
            print(f"ODS already at correct path: {OUTPUT}")

        # Remove temp xlsx
        if os.path.exists(tmp_xlsx):
            os.remove(tmp_xlsx)

        print(f"Initial ODS file created via LibreOffice conversion: {OUTPUT}")

    # Verify file exists
    if os.path.exists(OUTPUT):
        print(f"Verified: {OUTPUT} exists ({os.path.getsize(OUTPUT)} bytes)")
    else:
        print(f"ERROR: {OUTPUT} was not created!")

    # GUI-ready startup: open the file in LibreOffice Calc
    # Also open Chrome for web searches
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui('google-chrome --new-window "https://search.crossref.org/"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0")


create_initial()
