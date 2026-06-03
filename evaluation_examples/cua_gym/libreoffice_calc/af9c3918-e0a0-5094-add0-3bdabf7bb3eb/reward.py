"""
Reward Script: Individual Rep Forecast Sheet with SUMIFS, Weighted Value, and Gap formulas
Task ID: calc_sales_forecast_rep_individual_042
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): Weighted Value formula in E2:E50 (IF with Commit=0.9, BestCase=0.65, Pipeline=0.3)
  - Component 2 (0.30): SUMIFS formulas in H2:J13 for monthly Commit/BestCase/Pipeline totals
  - Component 3 (0.20): Gap formula in L2:L13 (=H+I+J-K)
  - Component 4 (0.20): Conditional formatting on L2:L13 (red if <0, green if >=0)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_forecast_rep_individual_042'
SHEET_NAME = 'RepForecast_Template'


def normalize_formula(formula):
    """Normalize formula for comparison: strip spaces, uppercase."""
    if formula is None:
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------------------------
    # Component 1: Weighted Value formula in E2:E50 (0.30 points)
    # Expected pattern: =IF(C2="Commit",B2*0.9,IF(C2="Best Case",B2*0.65,IF(C2="Pipeline",B2*0.3,0)))
    # This FAILS on initial (all E2:E50 are None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        e_formulas_ok = 0
        e_formulas_checked = 0

        for row in range(2, 51):  # rows 2 to 50
            cell_e = ws.cell(row=row, column=5)  # column E = 5
            val = cell_e.value

            if val is None:
                # No formula present — row is empty in initial, also valid in golden for empty rows
                # but the golden has E2:E50 all with formulas
                continue

            if not isinstance(val, str) or not val.startswith('='):
                continue

            e_formulas_checked += 1
            norm = normalize_formula(val)

            # Check for the required pattern:
            # =IF(Cx="Commit",Bx*0.9,IF(Cx="Best Case",Bx*0.65,IF(Cx="Pipeline",Bx*0.3,0)))
            # Must contain all three stages with their weights
            has_commit = 'COMMIT' in norm and '0.9' in norm
            has_bestcase = 'BESTCASE' in norm and '0.65' in norm
            has_pipeline = 'PIPELINE' in norm and '0.3' in norm
            is_if_formula = norm.startswith('=IF(')

            if is_if_formula and has_commit and has_bestcase and has_pipeline:
                e_formulas_ok += 1

        # Need at least 15 rows to have formulas (initial has 15 data rows, golden has all 49)
        # The golden file has formulas in E2:E50 (all 49 rows)
        if e_formulas_ok >= 49:
            print(f"PASS: Component 1 — Weighted Value formulas in E2:E50: {e_formulas_ok}/49 rows correct (0.30 pts)")
            total_score += 0.30
        elif e_formulas_ok >= 15:
            # Partial: at least data rows have formulas
            partial = 0.15
            print(f"PARTIAL: Component 1 — Weighted Value formulas: {e_formulas_ok}/49 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Weighted Value formulas: only {e_formulas_ok} valid formulas found in E2:E50")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: SUMIFS formulas in H2:J13 (0.30 points)
    # Expected:
    #   H2:H13 = SUMIFS summing B column where Stage='Commit' and Close Month=G row
    #   I2:I13 = SUMIFS for 'Best Case' by month
    #   J2:J13 = SUMIFS for 'Pipeline' by month
    # This FAILS on initial (all empty) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        h_ok = 0
        i_ok = 0
        j_ok = 0

        for row in range(2, 14):  # rows 2 to 13 (12 months)
            # Column H = 8 (Commit SUMIFS)
            h_val = ws.cell(row=row, column=8).value
            if h_val and isinstance(h_val, str) and h_val.startswith('='):
                norm_h = normalize_formula(h_val)
                if 'SUMIFS' in norm_h and 'COMMIT' in norm_h:
                    h_ok += 1

            # Column I = 9 (Best Case SUMIFS)
            i_val = ws.cell(row=row, column=9).value
            if i_val and isinstance(i_val, str) and i_val.startswith('='):
                norm_i = normalize_formula(i_val)
                if 'SUMIFS' in norm_i and 'BESTCASE' in norm_i:
                    i_ok += 1

            # Column J = 10 (Pipeline SUMIFS)
            j_val = ws.cell(row=row, column=10).value
            if j_val and isinstance(j_val, str) and j_val.startswith('='):
                norm_j = normalize_formula(j_val)
                if 'SUMIFS' in norm_j and 'PIPELINE' in norm_j:
                    j_ok += 1

        total_sumifs = h_ok + i_ok + j_ok
        if h_ok == 12 and i_ok == 12 and j_ok == 12:
            print(f"PASS: Component 2 — SUMIFS formulas in H2:J13: H={h_ok}/12, I={i_ok}/12, J={j_ok}/12 (0.30 pts)")
            total_score += 0.30
        elif total_sumifs >= 18:  # at least 50% of 36
            partial = 0.15
            print(f"PARTIAL: Component 2 — SUMIFS formulas partial: H={h_ok}, I={i_ok}, J={j_ok}/12 each ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — SUMIFS formulas missing: H={h_ok}/12, I={i_ok}/12, J={j_ok}/12")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Gap formula in L2:L13 (0.20 points)
    # Expected: =H2+I2+J2-K2 (and equivalently for rows 3-13)
    # This FAILS on initial (all empty) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        l_ok = 0

        for row in range(2, 14):  # rows 2 to 13
            l_val = ws.cell(row=row, column=12).value  # column L = 12
            if l_val and isinstance(l_val, str) and l_val.startswith('='):
                norm_l = normalize_formula(l_val)
                # Pattern: =Hx+Ix+Jx-Kx
                # Must contain H, I, J, K references and subtraction of K
                h_ref = f'H{row}'
                i_ref = f'I{row}'
                j_ref = f'J{row}'
                k_ref = f'K{row}'
                if (h_ref in norm_l and i_ref in norm_l and
                        j_ref in norm_l and k_ref in norm_l and
                        '-' in norm_l):
                    l_ok += 1

        if l_ok == 12:
            print(f"PASS: Component 3 — Gap formula in L2:L13: {l_ok}/12 rows correct (0.20 pts)")
            total_score += 0.20
        elif l_ok >= 6:
            partial = 0.10
            print(f"PARTIAL: Component 3 — Gap formula partial: {l_ok}/12 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Gap formula: only {l_ok}/12 correct formulas in L2:L13")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional formatting on L2:L13 (0.20 points)
    # Expected: red fill if value < 0, green fill if value >= 0
    # This FAILS on initial (no CF) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        cf_ranges = list(cf_rules)

        # Find CF applying to column L
        found_l_cf = False
        has_red_below_zero = False
        has_green_above_zero = False

        for cf in cf_ranges:
            cf_str = str(cf)
            if 'L' in cf_str:
                found_l_cf = True
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        formula_vals = getattr(rule, 'formula', [])
                        op = getattr(rule, 'operator', '')

                        # Check for red rule: lessThan 0
                        if op == 'lessThan' and formula_vals and formula_vals[0] == '0':
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                try:
                                    fg = rule.dxf.fill.fgColor.rgb
                                    # Red: FFFF0000 or similar (high red component)
                                    if fg and 'FF0000' in fg.upper():
                                        has_red_below_zero = True
                                except Exception:
                                    pass

                        # Check for green rule: greaterThanOrEqual 0
                        if op == 'greaterThanOrEqual' and formula_vals and formula_vals[0] == '0':
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                try:
                                    fg = rule.dxf.fill.fgColor.rgb
                                    # Green: FF00FF00 or similar (high green component)
                                    if fg and '00FF00' in fg.upper():
                                        has_green_above_zero = True
                                except Exception:
                                    pass

        if found_l_cf and has_red_below_zero and has_green_above_zero:
            print(f"PASS: Component 4 — Conditional formatting on L2:L13: red<0, green>=0 (0.20 pts)")
            total_score += 0.20
        elif found_l_cf and (has_red_below_zero or has_green_above_zero):
            partial = 0.10
            print(f"PARTIAL: Component 4 — Conditional formatting partial: red={has_red_below_zero}, green={has_green_above_zero} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Conditional formatting on L2:L13: found_CF={found_l_cf}, "
                  f"red_rule={has_red_below_zero}, green_rule={has_green_above_zero}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
