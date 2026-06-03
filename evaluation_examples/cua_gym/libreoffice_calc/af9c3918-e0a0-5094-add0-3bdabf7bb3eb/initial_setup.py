"""
Initial Setup: Individual rep forecast sheet - deals table with monthly summary
Task ID: calc_sales_forecast_rep_individual_042
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_forecast_rep_individual_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'RepForecast_Template'

    # ---- Section 1: Deal Entry Table (A1:E50) ----
    # Headers
    headers_section1 = ['Deal Name', 'Deal Size', 'Stage', 'Close Month', 'Weighted Value']
    for col, h in enumerate(headers_section1, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Realistic deal data
    deals = [
        ('Acme Corp CRM Expansion',     85000,  'Commit',    'Jan'),
        ('TechNova Platform License',   140000, 'Best Case', 'Jan'),
        ('Horizon Analytics Suite',     62000,  'Pipeline',  'Jan'),
        ('BlueSky Solutions Renewal',   95000,  'Commit',    'Feb'),
        ('PrecisionTech Upgrade',       55000,  'Best Case', 'Feb'),
        ('NorthStar Media Deal',        200000, 'Pipeline',  'Feb'),
        ('Apex Industries Integration', 120000, 'Commit',    'Mar'),
        ('Crestview Finance Tools',     78000,  'Best Case', 'Mar'),
        ('Summit Retail Package',       45000,  'Pipeline',  'Mar'),
        ('Evergreen Healthcare CRM',    160000, 'Commit',    'Apr'),
        ('Cascade Software Bundle',     88000,  'Best Case', 'Apr'),
        ('Prism Data Platform',         115000, 'Pipeline',  'Apr'),
        ('Titan Manufacturing ERP',     175000, 'Commit',    'May'),
        ('Stellar Communications',      92000,  'Best Case', 'May'),
        ('Nexus Cloud Migration',       67000,  'Pipeline',  'May'),
    ]

    for r, (name, size, stage, month) in enumerate(deals, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=size)
        ws.cell(row=r, column=3, value=stage)
        ws.cell(row=r, column=4, value=month)
        # Column E (Weighted Value) left EMPTY — task requires agent to add formula

    # Format column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    # ---- Section 2: Monthly Summary (G1:L14) ----
    # Headers
    summary_headers = ['Month', 'Commit', 'Best Case', 'Pipeline', 'Monthly Quota', 'Gap']
    for col, h in enumerate(summary_headers, 7):  # G=7, H=8, I=9, J=10, K=11, L=12
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Months Jan-Dec in G2:G13
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    for r, month in enumerate(months, 2):
        ws.cell(row=r, column=7, value=month)  # col G

    # Monthly quota $150,000 in K2:K13
    for r in range(2, 14):
        cell = ws.cell(row=r, column=11, value=150000)  # col K
        cell.number_format = '$#,##0'

    # H2:J13 (Commit/Best Case/Pipeline totals) left EMPTY — task requires SUMIFS
    # L2:L13 (Gap) left EMPTY — task requires Gap formula
    # No conditional formatting on L — task requires agent to add it

    # Format summary column widths
    ws.column_dimensions['F'].width = 3   # spacer
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14
    ws.column_dimensions['J'].width = 14
    ws.column_dimensions['K'].width = 16
    ws.column_dimensions['L'].width = 14

    # Format header row height
    ws.row_dimensions[1].height = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
