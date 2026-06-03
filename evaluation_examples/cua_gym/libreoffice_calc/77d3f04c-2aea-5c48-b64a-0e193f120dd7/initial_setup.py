"""
Initial Setup: Product sales trend - monthly units sold for top 5 products
Task ID: calc_sales_product_trend_018
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_product_trend_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: MonthlySales ---
    ws = wb.active
    ws.title = 'MonthlySales'

    # Headers: Month, Enterprise Suite, SMB Pack, Starter, Add-ons, Support Plans
    headers = ['Month', 'Enterprise Suite', 'SMB Pack', 'Starter', 'Add-ons', 'Support Plans']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Monthly data: Jan through Dec (12 rows)
    # Units sold per product, values in range 50-2400 as specified
    data = [
        ['Jan', 1820, 980, 2350, 340, 510],
        ['Feb', 1740, 1020, 2180, 290, 490],
        ['Mar', 1990, 1150, 2400, 410, 560],
        ['Apr', 2080, 1230, 2290, 380, 600],
        ['May', 2150, 1310, 2100, 440, 650],
        ['Jun', 2200, 1420, 1980, 470, 710],
        ['Jul', 1950, 1280, 1870, 390, 670],
        ['Aug', 1800, 1100, 1750, 350, 620],
        ['Sep', 2050, 1350, 2050, 430, 690],
        ['Oct', 2300, 1500, 2200, 510, 750],
        ['Nov', 2380, 1650, 2380, 580, 810],
        ['Dec', 2100, 1480, 2150, 490, 760],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets:', wb.sheetnames)
    print(f'Data rows: {ws.max_row - 1} (excluding header)')


create_initial()
