"""
Reward Script: Product Sales Trend Chart (Multi-Line)
Task ID: calc_sales_product_trend_018
Domain: libreoffice_calc

Task: Create a multi-line chart on a new sheet 'TrendChart' showing monthly units sold
for 5 products (Enterprise Suite, SMB Pack, Starter, Add-ons, Support Plans) over 12 months.
Chart title: 'Monthly Product Sales Trends', with 5 series, x-axis from MonthlySales!A2:A13,
y-axis units sold, and data markers visible on each data point.

Scoring Rubric:
  Component 1: 'TrendChart' sheet exists                          — 0.25 points
  Component 2: A LineChart with exactly 5 series exists           — 0.30 points
  Component 3: Chart title is 'Monthly Product Sales Trends'      — 0.20 points
  Component 4: Series reference correct data ranges (B-F, 2-13)   — 0.15 points
  Component 5: Data markers are visible on each series            — 0.10 points
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_product_trend_018'


def get_title_text(title_obj):
    """Extract text string from an openpyxl chart Title object."""
    try:
        if title_obj is None:
            return None
        # Navigate the rich text structure
        tx = title_obj.tx
        if tx is None:
            return None
        rich = tx.rich
        if rich is None:
            return None
        paragraphs = rich.p
        if not paragraphs:
            return None
        texts = []
        for para in paragraphs:
            for run in (para.r or []):
                if run.t:
                    texts.append(run.t)
        return ''.join(texts).strip() if texts else None
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'TrendChart' sheet exists (0.25 points)
    # This FAILS on initial file (only has MonthlySales), PASSES on golden
    try:
        if 'TrendChart' in wb.sheetnames:
            print("PASS: Component 1 — 'TrendChart' sheet exists (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — 'TrendChart' sheet not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Early exit if TrendChart doesn't exist — remaining components can't be checked
    if 'TrendChart' not in wb.sheetnames:
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws_chart = wb['TrendChart']

    # Component 2: A LineChart with exactly 5 series exists (0.30 points)
    # This FAILS on initial file (no TrendChart), PASSES on golden
    try:
        charts = ws_chart._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_type = type(chart).__name__
            num_series = len(chart.series)
            if chart_type == 'LineChart' and num_series == 5:
                print(f"PASS: Component 2 — LineChart with 5 series found (0.30 pts)")
                total_score += 0.30
            elif chart_type == 'LineChart' and num_series != 5:
                print(f"FAIL: Component 2 — LineChart exists but has {num_series} series (expected 5)")
            else:
                print(f"FAIL: Component 2 — Chart type is '{chart_type}' (expected LineChart)")
        else:
            print(f"FAIL: Component 2 — No charts found on 'TrendChart' sheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Chart title is 'Monthly Product Sales Trends' (0.20 points)
    # This FAILS on initial file, PASSES on golden
    try:
        charts = ws_chart._charts
        if len(charts) >= 1:
            chart = charts[0]
            title_text = get_title_text(chart.title)
            if title_text and title_text.strip() == 'Monthly Product Sales Trends':
                print(f"PASS: Component 3 — Chart title is 'Monthly Product Sales Trends' (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — Expected title 'Monthly Product Sales Trends', found: {repr(title_text)}")
        else:
            print(f"FAIL: Component 3 — No charts available to check title")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Series reference correct data ranges from MonthlySales (0.15 points)
    # Each series should reference MonthlySales columns B-F rows 2-13 for values
    # and MonthlySales column A rows 2-13 for categories
    # This FAILS on initial file, PASSES on golden
    try:
        charts = ws_chart._charts
        if len(charts) >= 1 and len(charts[0].series) == 5:
            chart = charts[0]
            expected_val_cols = ['$B$', '$C$', '$D$', '$E$', '$F$']
            all_refs_ok = True
            for i, series in enumerate(chart.series):
                # Check value reference
                val_ref = None
                cat_ref = None
                if hasattr(series, 'val') and series.val and hasattr(series.val, 'numRef') and series.val.numRef:
                    val_ref = series.val.numRef.f
                if hasattr(series, 'cat') and series.cat:
                    if hasattr(series.cat, 'numRef') and series.cat.numRef:
                        cat_ref = series.cat.numRef.f
                    elif hasattr(series.cat, 'strRef') and series.cat.strRef:
                        cat_ref = series.cat.strRef.f

                col_ok = val_ref and expected_val_cols[i] in val_ref and 'MonthlySales' in val_ref
                rows_ok = val_ref and '$2:' in val_ref and '$13' in val_ref
                cat_ok = cat_ref and 'MonthlySales' in cat_ref and '$A$' in cat_ref

                if not (col_ok and rows_ok):
                    print(f"FAIL: Component 4 — Series {i} val ref wrong: {repr(val_ref)}")
                    all_refs_ok = False
                    break
                if not cat_ok:
                    print(f"FAIL: Component 4 — Series {i} cat ref wrong: {repr(cat_ref)}")
                    all_refs_ok = False
                    break

            if all_refs_ok:
                print(f"PASS: Component 4 — All 5 series reference correct data ranges from MonthlySales (0.15 pts)")
                total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Cannot check data references (chart or series count issue)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Data markers are visible on each series (0.10 points)
    # Each series should have a marker symbol set (not 'none' or missing)
    # This FAILS on initial file, PASSES on golden
    try:
        charts = ws_chart._charts
        if len(charts) >= 1 and len(charts[0].series) == 5:
            chart = charts[0]
            all_markers_ok = True
            for i, series in enumerate(chart.series):
                marker_ok = False
                if hasattr(series, 'marker') and series.marker is not None:
                    sym = series.marker.symbol
                    # symbol should be something other than None, 'none', or 'auto' without explicit set
                    if sym is not None and sym != 'none':
                        marker_ok = True
                if not marker_ok:
                    print(f"FAIL: Component 5 — Series {i} has no visible marker (symbol={getattr(getattr(series, 'marker', None), 'symbol', None)})")
                    all_markers_ok = False
                    break

            if all_markers_ok:
                print(f"PASS: Component 5 — All 5 series have visible data markers (0.10 pts)")
                total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Cannot check markers (chart or series count issue)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
