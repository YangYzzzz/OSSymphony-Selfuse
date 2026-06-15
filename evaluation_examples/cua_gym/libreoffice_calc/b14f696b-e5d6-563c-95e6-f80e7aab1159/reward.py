"""
Reward Script: Multi-period comparative financial dashboard with variance analysis,
               conditional formatting, and a chart.
Task ID: calc_gpm_010
Domain: libreoffice_calc
Scoring:
  Component 1: Variance formulas in D4:D9 and G4:G9           — 0.30 pts
  Component 2: Variance columns formatted as percentage        — 0.10 pts
  Component 3: Conditional formatting on variance columns      — 0.20 pts
  Component 4: Data bars on B4:B9 and E4:E9                   — 0.15 pts
  Component 5: Grouped bar chart with correct title            — 0.15 pts
  Component 6: Thick borders around quarter sections           — 0.10 pts
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_010'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: FinDash sheet must exist
    if 'FinDash' not in wb.sheetnames:
        print(f"FAIL: Sheet 'FinDash' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['FinDash']

    # =========================================================================
    # Component 1: Variance formulas in D4:D9 and G4:G9 (0.30 points)
    # Initial has None in these cells; golden has =(Bx-Cx)/Cx and =(Ex-Fx)/Fx
    # =========================================================================
    try:
        formula_score = 0.0
        d_formulas_ok = 0
        g_formulas_ok = 0

        for row in range(4, 10):
            # Check D column: =(Bx-Cx)/Cx
            d_val = ws.cell(row=row, column=4).value
            if d_val is not None and isinstance(d_val, str):
                normalized = d_val.upper().replace(" ", "")
                expected = f"=(B{row}-C{row})/C{row}".upper()
                if normalized == expected:
                    d_formulas_ok += 1

            # Check G column: =(Ex-Fx)/Fx
            g_val = ws.cell(row=row, column=7).value
            if g_val is not None and isinstance(g_val, str):
                normalized = g_val.upper().replace(" ", "")
                expected = f"=(E{row}-F{row})/F{row}".upper()
                if normalized == expected:
                    g_formulas_ok += 1

        total_formulas = d_formulas_ok + g_formulas_ok
        if total_formulas == 12:
            print(f"PASS: Component 1 — All 12 variance formulas present (0.30 pts)")
            total_score += 0.30
        elif total_formulas >= 6:
            print(f"PARTIAL: Component 1 — {total_formulas}/12 variance formulas (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {total_formulas}/12 variance formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Variance columns formatted as percentage (0.10 points)
    # Initial has 'General' format; golden has '0.00%'
    # =========================================================================
    try:
        pct_count = 0
        for row in range(4, 10):
            d_nf = ws.cell(row=row, column=4).number_format
            g_nf = ws.cell(row=row, column=7).number_format
            if d_nf and '%' in str(d_nf):
                pct_count += 1
            if g_nf and '%' in str(g_nf):
                pct_count += 1

        if pct_count >= 10:
            print(f"PASS: Component 2 — {pct_count}/12 variance cells formatted as percentage (0.10 pts)")
            total_score += 0.10
        elif pct_count >= 6:
            print(f"PARTIAL: Component 2 — {pct_count}/12 variance cells formatted as percentage (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — Only {pct_count}/12 variance cells have percentage format")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Conditional formatting on variance columns (0.20 points)
    # Initial has 0 CF rules; golden has CF rules on D4:D9 and G4:G9
    # with green font for >0 and red font for <0
    # =========================================================================
    try:
        cf_ranges_with_cellis = set()
        green_rule_count = 0
        red_rule_count = 0

        for cf in ws.conditional_formatting:
            range_str = str(cf)
            for rule in cf.rules:
                if rule.type == 'cellIs':
                    cf_ranges_with_cellis.add(range_str)
                    # Count green (positive) rules
                    if rule.operator in ('greaterThan', 'greaterThanOrEqual'):
                        green_rule_count += 1
                    # Count red (negative) rules
                    if rule.operator in ('lessThan', 'lessThanOrEqual'):
                        red_rule_count += 1

        # Check that CF is applied to variance columns (D and G)
        covers_d = any('D' in r for r in cf_ranges_with_cellis)
        covers_g = any('G' in r for r in cf_ranges_with_cellis)

        if covers_d and covers_g and green_rule_count > 0 and red_rule_count > 0:
            print(f"PASS: Component 3 — Conditional formatting with green/red on D and G columns (0.20 pts)")
            total_score += 0.20
        elif (covers_d or covers_g) and (green_rule_count > 0 or red_rule_count > 0):
            print(f"PARTIAL: Component 3 — Partial conditional formatting (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting on variance columns. CF ranges: {cf_ranges_with_cellis}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Data bars on B4:B9 and E4:E9 (0.15 points)
    # Initial has 0 data bar rules; golden has data bars on both ranges
    # =========================================================================
    try:
        databar_ranges = set()
        for cf in ws.conditional_formatting:
            range_str = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    databar_ranges.add(range_str)

        covers_b = any('B' in r for r in databar_ranges)
        covers_e = any('E' in r for r in databar_ranges)

        if covers_b and covers_e:
            print(f"PASS: Component 4 — Data bars on B and E actual columns (0.15 pts)")
            total_score += 0.15
        elif covers_b or covers_e:
            print(f"PARTIAL: Component 4 — Data bars on one actual column only (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 4 — No data bars found. DataBar ranges: {databar_ranges}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Grouped bar chart with correct title (0.15 points)
    # Initial has 0 charts; golden has 1 chart titled "Q1 vs Q2 Performance"
    # =========================================================================
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_type = getattr(chart, 'type', None)
            chart_grouping = getattr(chart, 'grouping', None)

            # Extract chart title text
            chart_title_text = None
            if chart.title:
                try:
                    # Navigate the title object to find text
                    if hasattr(chart.title, 'tx') and chart.title.tx:
                        if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                            for p in chart.title.tx.rich.p:
                                for r in p.r:
                                    if r.t:
                                        chart_title_text = r.t
                except Exception:
                    pass

            is_bar_type = chart_type in ('col', 'bar')
            title_matches = chart_title_text and 'q1' in chart_title_text.lower() and 'q2' in chart_title_text.lower()
            has_multiple_series = len(chart.series) >= 2

            if is_bar_type and title_matches and has_multiple_series:
                print(f"PASS: Component 5 — Grouped bar chart '{chart_title_text}' with {len(chart.series)} series (0.15 pts)")
                total_score += 0.15
            elif is_bar_type and has_multiple_series:
                print(f"PARTIAL: Component 5 — Bar chart present with {len(chart.series)} series but title mismatch (0.10 pts)")
                total_score += 0.10
            elif len(charts) >= 1:
                print(f"PARTIAL: Component 5 — Chart exists but type={chart_type}, series={len(chart.series)}, title='{chart_title_text}' (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No charts found in sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Thick borders around quarter sections (0.10 points)
    # Initial has no thick borders; golden has thick borders on section boundaries
    # Specifically: B column left=thick, D column right=thick, E column left=thick, G column right=thick
    # =========================================================================
    try:
        thick_border_count = 0
        # Check a sample of cells for thick borders
        # Q1 section: B3:D9 should have thick left on B and thick right on D
        # Q2 section: E3:G9 should have thick left on E and thick right on G (or similar)
        for row in range(3, 10):
            b_cell = ws.cell(row=row, column=2)
            if b_cell.border and b_cell.border.left and b_cell.border.left.style == 'thick':
                thick_border_count += 1

            d_cell = ws.cell(row=row, column=4)
            if d_cell.border and d_cell.border.right and d_cell.border.right.style == 'thick':
                thick_border_count += 1

            e_cell = ws.cell(row=row, column=5)
            if e_cell.border and e_cell.border.left and e_cell.border.left.style == 'thick':
                thick_border_count += 1

            g_cell = ws.cell(row=row, column=7)
            if g_cell.border and g_cell.border.right and g_cell.border.right.style == 'thick':
                thick_border_count += 1

        # Also check top/bottom thick borders on row 3 and row 9
        for col in range(2, 8):
            top_cell = ws.cell(row=3, column=col)
            if top_cell.border and top_cell.border.top and top_cell.border.top.style == 'thick':
                thick_border_count += 1
            bot_cell = ws.cell(row=9, column=col)
            if bot_cell.border and bot_cell.border.bottom and bot_cell.border.bottom.style == 'thick':
                thick_border_count += 1

        if thick_border_count >= 8:
            print(f"PASS: Component 6 — Thick borders found ({thick_border_count} thick border sides detected) (0.10 pts)")
            total_score += 0.10
        elif thick_border_count >= 4:
            print(f"PARTIAL: Component 6 — Some thick borders found ({thick_border_count} sides) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Insufficient thick borders ({thick_border_count} sides found)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
