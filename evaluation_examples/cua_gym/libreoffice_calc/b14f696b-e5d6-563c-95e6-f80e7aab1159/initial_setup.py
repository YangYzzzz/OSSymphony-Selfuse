"""
Initial Setup: Multi-period comparative financial dashboard
Task ID: calc_gpm_010
Domain: libreoffice_calc

Creates FinDash sheet with Q1/Q2 actual and budget data.
Variance columns (D, G) are LEFT EMPTY — filling them is the agent's task.
No conditional formatting, no data bars, no chart, no section borders.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FinDash'

    # --- Row 1: Merged title ---
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = 'Quarterly Financial Dashboard - FY2025'
    title_cell.font = Font(size=16, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='FF000080', end_color='FF000080', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 3: Headers ---
    headers = ['Category', 'Q1 Actual', 'Q1 Budget', 'Q1 Var%',
               'Q2 Actual', 'Q2 Budget', 'Q2 Var%']
    navy_fill = PatternFill(start_color='FF000080', end_color='FF000080', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = header_align

    # --- Rows 4-9: Financial data ---
    # Format: Category, Q1 Actual, Q1 Budget, Q1 Var%, Q2 Actual, Q2 Budget, Q2 Var%
    # Var% columns (D, G) are left EMPTY — that's the task
    data = [
        ['Revenue',       520000, 500000, None, 540000, 525000, None],
        ['COGS',          210000, 200000, None, 225000, 210000, None],
        ['Gross Profit',  310000, 300000, None, 315000, 315000, None],
        ['Operating Exp', 180000, 175000, None, 185000, 180000, None],
        ['EBITDA',        130000, 125000, None, 130000, 135000, None],
        ['Net Income',     95000,  90000, None,  98000, 100000, None],
    ]

    for r, row_data in enumerate(data, 4):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # --- Currency formatting on B:C and E:F ---
    for row in range(4, 10):
        for col in [2, 3, 5, 6]:  # B, C, E, F
            ws.cell(row=row, column=col).number_format = '$#,##0'

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 14

    # --- Row height for title ---
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
