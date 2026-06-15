"""
Reward Script: Sort quarterly earnings by date and create EPS line chart
Task ID: osworld_calc_sort_date_chart_010
Domain: libreoffice_calc
Scoring:
  Component 1: Data rows sorted by Announcement Date ascending (0.5 pts)
  Component 2: Line chart exists with title 'EPS Over Time' (0.3 pts)
  Component 3: Chart series references EPS column (E) and categories reference date column (A) (0.2 pts)
Total: 1.0
"""

import os
import openpyxl
from openpyxl.chart import LineChart

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sort_date_chart_010'


def get_chart_title_text(chart):
    """Extract the title string from a chart title object."""
    try:
        title_obj = chart.title
        if title_obj is None:
            return None
        # Navigate through the openpyxl title structure
        tx = title_obj.tx
        if tx is None:
            return None
        rich = tx.rich
        if rich is None:
            return None
        paragraphs = rich.p
        if not paragraphs:
            return None
        parts = []
        for para in paragraphs:
            for run in (para.r or []):
                if run.t:
                    parts.append(run.t)
        return ''.join(parts).strip()
    except Exception:
        return None


def get_axis_title_text(axis_title_obj):
    """Extract the title string from an axis title object."""
    try:
        if axis_title_obj is None:
            return None
        tx = axis_title_obj.tx
        if tx is None:
            return None
        rich = tx.rich
        if rich is None:
            return None
        paragraphs = rich.p
        if not paragraphs:
            return None
        parts = []
        for para in paragraphs:
            for run in (para.r or []):
                if run.t:
                    parts.append(run.t)
        return ''.join(parts).strip()
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Earnings' sheet must exist
    if 'Earnings' not in wb.sheetnames:
        print("CRITICAL: 'Earnings' sheet not found. Cannot verify task.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Earnings']

    # Component 1: Data rows sorted by Announcement Date ascending (0.5 points)
    # The initial file has dates in random order; the task requires ascending sort.
    # We check that dates in column A (rows 2 onward) form a non-decreasing sequence.
    try:
        date_values = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
            cell = row[0]
            if cell.value is not None:
                date_values.append(cell.value)

        if len(date_values) >= 2:
            is_sorted_asc = all(
                date_values[i] <= date_values[i + 1]
                for i in range(len(date_values) - 1)
            )
            if is_sorted_asc:
                print(f"PASS: Component 1 — {len(date_values)} date rows are sorted ascending "
                      f"(first: {date_values[0]}, last: {date_values[-1]}) (0.5 pts)")
                total_score += 0.5
            else:
                # Find the first out-of-order pair for diagnostic purposes
                for i in range(len(date_values) - 1):
                    if date_values[i] > date_values[i + 1]:
                        print(f"FAIL: Component 1 — dates NOT sorted ascending: "
                              f"row {i+2} ({date_values[i]}) > row {i+3} ({date_values[i+1]})")
                        break
        else:
            print(f"FAIL: Component 1 — not enough date rows found (found {len(date_values)})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A line chart titled 'EPS Over Time' exists (0.3 points)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, LineChart)]

        if len(charts) == 0:
            print("FAIL: Component 2 — no charts found in 'Earnings' sheet")
        elif len(line_charts) == 0:
            print(f"FAIL: Component 2 — {len(charts)} chart(s) found but none are line charts "
                  f"(types: {[type(c).__name__ for c in charts]})")
        else:
            # Check if any line chart has the title 'EPS Over Time'
            matched_chart = None
            for chart in line_charts:
                title_text = get_chart_title_text(chart)
                if title_text and title_text.strip() == 'EPS Over Time':
                    matched_chart = chart
                    break

            if matched_chart is not None:
                print(f"PASS: Component 2 — line chart with title 'EPS Over Time' found (0.3 pts)")
                total_score += 0.3
            else:
                found_titles = [get_chart_title_text(c) for c in line_charts]
                print(f"FAIL: Component 2 — line chart(s) found but title does not match. "
                      f"Found titles: {found_titles}; expected 'EPS Over Time'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Chart series references EPS column (E) data and
    #              categories reference Announcement Date column (A) (0.2 points)
    # This ensures the chart shows EPS over time, not some other column.
    try:
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, LineChart)]

        # Find the titled chart (or first line chart)
        target_chart = None
        for chart in line_charts:
            title_text = get_chart_title_text(chart)
            if title_text and title_text.strip() == 'EPS Over Time':
                target_chart = chart
                break
        if target_chart is None and line_charts:
            target_chart = line_charts[0]

        if target_chart is not None and len(target_chart.series) > 0:
            series = target_chart.series[0]

            # Check y-axis data references EPS column (column E)
            eps_series_formula = None
            try:
                val_ref = series.val
                if val_ref and val_ref.numRef:
                    formula = val_ref.numRef.f
                    if formula:
                        eps_series_formula = formula
                        print(f"  Series val ref: {formula}")
            except Exception as e:
                print(f"  Series val ref check error: {e}")

            eps_in_series = (
                eps_series_formula is not None and
                ('$E$' in eps_series_formula or '!E' in eps_series_formula)
            )

            # Check x-axis categories reference date column (column A)
            cat_formula = None
            try:
                cat = series.cat
                if cat is not None:
                    if cat.numRef:
                        cat_formula = cat.numRef.f
                    elif cat.strRef:
                        cat_formula = cat.strRef.f
                    if cat_formula:
                        print(f"  Categories ref: {cat_formula}")
            except Exception as e:
                print(f"  Categories check error: {e}")

            date_in_categories = (
                cat_formula is not None and
                ('$A$' in cat_formula or '!A' in cat_formula)
            )

            if eps_in_series and date_in_categories:
                print(f"PASS: Component 3 — chart series uses EPS (col E) as y-axis and "
                      f"dates (col A) as x-axis categories (0.2 pts)")
                total_score += 0.2
            elif eps_in_series:
                print(f"FAIL: Component 3 — EPS data correct but categories do not reference "
                      f"column A (date column)")
            elif date_in_categories:
                print(f"FAIL: Component 3 — date categories correct but series does not reference "
                      f"EPS column E")
            else:
                print(f"FAIL: Component 3 — series does not reference EPS (col E) or "
                      f"dates (col A)")
        else:
            print("FAIL: Component 3 — no line chart with series found to verify data references")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
