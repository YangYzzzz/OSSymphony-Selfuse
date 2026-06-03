"""
Initial Setup: Sort quarterly earnings announcements by date and create line chart
Task ID: osworld_calc_sort_date_chart_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from datetime import date

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_sort_date_chart_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Earnings ---
    ws = wb.active
    ws.title = 'Earnings'

    # Headers
    headers = ['Announcement Date', 'Company', 'Revenue', 'Net Income', 'EPS']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data in RANDOM ORDER (NOT sorted by date) — 12 quarterly records
    # Columns: Announcement Date, Company, Revenue ($M), Net Income ($M), EPS ($)
    data = [
        [date(2024, 7, 25), 'Nexlium Corp',       4821.3, 612.7,  2.14],
        [date(2023, 1, 19), 'Nexlium Corp',        3920.5, 498.2,  1.74],
        [date(2024, 10, 24), 'Nexlium Corp',       5103.8, 687.4,  2.41],
        [date(2023, 7, 20), 'Nexlium Corp',        4215.6, 553.1,  1.94],
        [date(2023, 4, 20), 'Nexlium Corp',        4047.9, 521.4,  1.83],
        [date(2024, 1, 18), 'Nexlium Corp',        4408.2, 574.3,  2.01],
        [date(2023, 10, 26), 'Nexlium Corp',       4332.7, 562.9,  1.97],
        [date(2024, 4, 25), 'Nexlium Corp',        4659.5, 595.8,  2.08],
        [date(2022, 10, 27), 'Nexlium Corp',       3744.2, 462.5,  1.62],
        [date(2022, 7, 21),  'Nexlium Corp',       3605.8, 441.9,  1.55],
        [date(2022, 4, 21),  'Nexlium Corp',       3489.1, 428.6,  1.50],
        [date(2022, 1, 20),  'Nexlium Corp',       3312.4, 398.7,  1.40],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)
        # Format date column
        ws.cell(row=r, column=1).number_format = 'yyyy-mm-dd'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
