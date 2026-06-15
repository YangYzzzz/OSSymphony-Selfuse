"""
Reward Script: MergeDuplicates macro — scan column A for duplicates, keep first occurrence, delete other rows
Task ID: calc_mcp_022
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Macro 'MergeDuplicates' exists in LibreOffice user macros
  Component 2 (0.4): All emails in column A are unique (no duplicates remain)
  Component 3 (0.3): First-occurrence data (Name, Phone) is preserved correctly
"""

import os
import glob as globmod

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_022'

# Known first-occurrence data from the initial file (row 2..row N before dedup)
# These are the Name and Phone values for the FIRST occurrence of each duplicated email
FIRST_OCCURRENCE_CHECKS = {
    'sarah.chen@techcorp.com': ('Sarah Chen', '(415) 555-0142'),
    'marcus.johnson@globex.net': ('Marcus Johnson', '(312) 555-0198'),
    'elena.rodriguez@innovate.io': ('Elena Rodriguez', '(646) 555-0233'),
    'priya.patel@nexuslab.org': ('Priya Patel', '(408) 555-0451'),
    'olivia.thompson@brightpath.co': ('Olivia Thompson', '(503) 555-0528'),
    'daniel.kim@quantumleap.io': ('Daniel Kim', '(213) 555-0614'),
    'rachel.green@freshstart.com': ('Rachel Green', '(617) 555-0742'),
    'mei.huang@pacificrim.co': ('Mei Huang', '(415) 555-1071'),
}


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # =========================================================================
    # Component 1: Macro 'MergeDuplicates' exists in LibreOffice user macros (0.3 pts)
    # =========================================================================
    try:
        # The macro should be stored as a .xba file in the LibreOffice user Basic directory
        macro_search_paths = [
            os.path.join(WORKDIR, '.config', 'libreoffice', 'user', 'basic', 'Standard', 'MergeDuplicates.xba'),
            os.path.join(WORKDIR, '.config', 'libreoffice', '4', 'user', 'basic', 'Standard', 'MergeDuplicates.xba'),
        ]
        # Also search with glob for any location
        glob_results = globmod.glob(
            os.path.join(WORKDIR, '.config', 'libreoffice', '*', 'user', 'basic', '**', 'MergeDuplicates.xba'),
            recursive=True
        )

        macro_found = False
        macro_has_sub = False
        macro_path = None

        # Check explicit paths first, then glob results
        all_paths = list(set(macro_search_paths + glob_results))
        for path in all_paths:
            if os.path.isfile(path):
                macro_found = True
                macro_path = path
                break

        if macro_found and macro_path:
            # Verify the file contains a Sub named MergeDuplicates
            content = open(macro_path, 'r', errors='ignore').read()
            if 'Sub MergeDuplicates' in content or 'sub MergeDuplicates' in content:
                macro_has_sub = True

        if macro_found and macro_has_sub:
            print(f"PASS: Component 1 -- Macro 'MergeDuplicates' found at {macro_path} with correct Sub definition (0.3 pts)")
            total_score += 0.3
        elif macro_found:
            print(f"FAIL: Component 1 -- Macro file found at {macro_path} but does not contain 'Sub MergeDuplicates'")
        else:
            print(f"FAIL: Component 1 -- Macro 'MergeDuplicates.xba' not found in LibreOffice user basic directories")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Load the spreadsheet for Components 2 and 3
    # =========================================================================
    file_path = os.path.join(WORKDIR, f'{TASK_ID}.xlsx')
    if not os.path.exists(file_path):
        print(f"CRITICAL: File not found: {file_path}")
        print(f"REWARD: {total_score}")
        return total_score

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print(f"REWARD: {total_score}")
        return total_score

    # Get the Contacts sheet
    if 'Contacts' in wb.sheetnames:
        ws = wb['Contacts']
    else:
        ws = wb.active
        print(f"WARNING: 'Contacts' sheet not found, using active sheet: {ws.title}")

    # Read all data rows
    data_rows = []
    for r in range(2, ws.max_row + 1):
        email = ws.cell(row=r, column=1).value
        name = ws.cell(row=r, column=2).value
        phone = ws.cell(row=r, column=3).value
        if email is not None:
            data_rows.append((email, name, phone))

    # =========================================================================
    # Component 2: All emails in column A are unique — no duplicates remain (0.4 pts)
    # =========================================================================
    try:
        emails = [row[0] for row in data_rows]
        email_set = set(emails)
        num_total = len(emails)
        num_unique = len(email_set)

        if num_total == num_unique and num_total > 0:
            print(f"PASS: Component 2 -- All {num_total} emails are unique, no duplicates (0.4 pts)")
            total_score += 0.4
        else:
            dup_count = num_total - num_unique
            print(f"FAIL: Component 2 -- Found {dup_count} duplicate email(s) among {num_total} rows")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Duplicates removed AND first-occurrence data preserved (0.3 pts)
    # This component requires that duplicate rows are gone (row count == unique count)
    # AND the kept rows have the correct Name/Phone from first occurrences.
    # This ensures it FAILS on initial_env (which still has duplicates).
    # =========================================================================
    try:
        emails = [row[0] for row in data_rows]
        has_no_duplicates = len(emails) == len(set(emails)) and len(emails) > 0

        if not has_no_duplicates:
            print(f"FAIL: Component 3 -- Cannot verify first-occurrence data: duplicates still present ({len(emails)} rows, {len(set(emails))} unique)")
        else:
            # Build a lookup from the current data (already deduplicated)
            email_to_data = {}
            for email, name, phone in data_rows:
                if email and email not in email_to_data:
                    email_to_data[email] = (name, phone)

            checks_passed = 0
            checks_total = len(FIRST_OCCURRENCE_CHECKS)

            for email, (expected_name, expected_phone) in FIRST_OCCURRENCE_CHECKS.items():
                if email in email_to_data:
                    actual_name, actual_phone = email_to_data[email]
                    name_ok = str(actual_name).strip() == expected_name.strip() if actual_name else False
                    phone_ok = str(actual_phone).strip() == expected_phone.strip() if actual_phone else False
                    if name_ok and phone_ok:
                        checks_passed += 1
                    else:
                        print(f"  INFO: {email} -- Name: expected '{expected_name}', got '{actual_name}'; Phone: expected '{expected_phone}', got '{actual_phone}'")
                else:
                    print(f"  INFO: {email} -- email not found in data (row may have been incorrectly deleted)")

            if checks_passed == checks_total and checks_total > 0:
                print(f"PASS: Component 3 -- All {checks_total} first-occurrence records preserved with correct Name and Phone after dedup (0.3 pts)")
                total_score += 0.3
            elif checks_passed > 0:
                partial = round(0.3 * (checks_passed / checks_total), 2)
                print(f"PARTIAL: Component 3 -- {checks_passed}/{checks_total} first-occurrence records correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 -- 0/{checks_total} first-occurrence records have correct data")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point — persist app state then verify
def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")

persist_app_state("libreoffice_calc")
verify_task()
