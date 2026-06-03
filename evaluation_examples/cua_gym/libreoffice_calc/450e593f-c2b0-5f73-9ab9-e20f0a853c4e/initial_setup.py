"""
Initial Setup: Write a macro called 'MergeDuplicates' that deduplicates column A
Task ID: calc_mcp_022
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    # Headers
    ws['A1'] = 'Email'
    ws['B1'] = 'Name'
    ws['C1'] = 'Phone'

    # Realistic contact data with deliberate duplicates
    # Format: (email, name, phone)
    contacts = [
        ('sarah.chen@techcorp.com', 'Sarah Chen', '(415) 555-0142'),
        ('marcus.johnson@globex.net', 'Marcus Johnson', '(312) 555-0198'),
        ('elena.rodriguez@innovate.io', 'Elena Rodriguez', '(646) 555-0233'),
        ('sarah.chen@techcorp.com', 'Sarah Chen (Dup)', '(415) 555-9999'),
        ('james.wilson@dataflow.com', 'James Wilson', '(206) 555-0317'),
        ('priya.patel@nexuslab.org', 'Priya Patel', '(408) 555-0451'),
        ('marcus.johnson@globex.net', 'Marcus Johnson (Dup)', '(312) 555-8888'),
        ('olivia.thompson@brightpath.co', 'Olivia Thompson', '(503) 555-0528'),
        ('daniel.kim@quantumleap.io', 'Daniel Kim', '(213) 555-0614'),
        ('rachel.green@freshstart.com', 'Rachel Green', '(617) 555-0742'),
        ('elena.rodriguez@innovate.io', 'Elena Rodriguez (Dup)', '(646) 555-7777'),
        ('alex.morgan@cloudpeak.net', 'Alex Morgan', '(720) 555-0855'),
        ('sarah.chen@techcorp.com', 'Sarah Chen (Dup2)', '(415) 555-6666'),
        ('liam.oshea@tradecraft.com', 'Liam O\'Shea', '(857) 555-0963'),
        ('mei.huang@pacificrim.co', 'Mei Huang', '(415) 555-1071'),
        ('carlos.mendez@solarbright.org', 'Carlos Mendez', '(305) 555-1189'),
        ('priya.patel@nexuslab.org', 'Priya Patel (Dup)', '(408) 555-5555'),
        ('natasha.volkov@eurasia.net', 'Natasha Volkov', '(212) 555-1297'),
        ('jordan.hayes@speedline.io', 'Jordan Hayes', '(469) 555-1385'),
        ('olivia.thompson@brightpath.co', 'Olivia Thompson (Dup)', '(503) 555-4444'),
        ('ben.carter@westwind.com', 'Ben Carter', '(602) 555-1422'),
        ('aisha.johnson@uplift.org', 'Aisha Johnson', '(773) 555-1536'),
        ('daniel.kim@quantumleap.io', 'Daniel Kim (Dup)', '(213) 555-3333'),
        ('sophie.laurent@artisan.fr', 'Sophie Laurent', '(347) 555-1648'),
        ('marcus.johnson@globex.net', 'Marcus Johnson (Dup2)', '(312) 555-2222'),
        ('tom.nakamura@eastgate.jp', 'Tom Nakamura', '(510) 555-1759'),
        ('rachel.green@freshstart.com', 'Rachel Green (Dup)', '(617) 555-1111'),
        ('isabella.rossi@bellavita.it', 'Isabella Rossi', '(917) 555-1867'),
        ('david.oconnell@summit.com', 'David O\'Connell', '(404) 555-1975'),
        ('mei.huang@pacificrim.co', 'Mei Huang (Dup)', '(415) 555-0000'),
    ]

    for r, (email, name, phone) in enumerate(contacts, 2):
        ws.cell(row=r, column=1, value=email)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=phone)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
