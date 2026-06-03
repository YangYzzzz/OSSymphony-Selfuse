"""
Reward Script: Pivot table drill-down on Electronics category
Task ID: calc_pivot_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Detail sheet for Electronics exists
  Component 2 (0.15): Detail sheet has correct headers (ID, Category, Product, Amount)
  Component 3 (0.25): All data rows in detail sheet have Category=Electronics
  Component 4 (0.15): Detail sheet has approximately correct row count (~55 rows)
  Component 5 (0.15): Sum of Amount in detail sheet equals ~67500
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_047'

def find_detail_sheet(wb, original_sheets):
    """
    Find the new detail sheet created by drilling down.
    It should be a sheet that did NOT exist in the original workbook.
    The expected name is 'Electronics' or similar, but we also check
    for any new sheet beyond the original two (SalesRecords, Summary).
    """
    original_names = {'SalesRecords', 'Summary'}
    new_sheets = [name for name in wb.sheetnames if name not in original_names]
    if not new_sheets:
        return None
    # Prefer a sheet with "Electronics" in the name
    for name in new_sheets:
        if 'electronics' in name.lower():
            return name
    # Otherwise return the first new sheet (drill-down may name it differently)
    return new_sheets[0]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Detail sheet for Electronics exists (0.30 points)
    # This is the PRIMARY task-introduced change: a new sheet should exist
    # beyond the original SalesRecords and Summary sheets.
    try:
        detail_sheet_name = find_detail_sheet(wb, {'SalesRecords', 'Summary'})
        if detail_sheet_name is not None:
            print(f"PASS: Component 1 - Detail sheet '{detail_sheet_name}' exists (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 - No new detail sheet found. Sheets: {wb.sheetnames}")
            # No detail sheet means no further checks are possible
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb[detail_sheet_name]

    # Component 2: Correct headers (0.15 points)
    # The detail sheet should mirror the source data columns: ID, Category, Product, Amount
    try:
        expected_headers = ['ID', 'Category', 'Product', 'Amount']
        actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        if actual_headers == expected_headers:
            print(f"PASS: Component 2 - Headers match {expected_headers} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 - Expected headers {expected_headers}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: All data rows have Category=Electronics (0.25 points)
    # This verifies the drill-down correctly filtered to Electronics only
    try:
        data_rows = 0
        electronics_rows = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2, values_only=True):
            val = row[0]
            if val is not None:
                data_rows += 1
                if str(val).strip() == 'Electronics':
                    electronics_rows += 1
        if data_rows > 0 and electronics_rows == data_rows:
            print(f"PASS: Component 3 - All {data_rows} data rows are Electronics (0.25 pts)")
            total_score += 0.25
        elif data_rows == 0:
            print(f"FAIL: Component 3 - No data rows found in detail sheet")
        else:
            print(f"FAIL: Component 3 - {electronics_rows}/{data_rows} rows are Electronics")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Approximate row count (~55 data rows) (0.15 points)
    # The task context says approximately 55 rows. Allow a range of 40-70 for tolerance.
    try:
        data_row_count = 0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if any(v is not None for v in row):
                data_row_count += 1
        if 40 <= data_row_count <= 70:
            print(f"PASS: Component 4 - Detail sheet has {data_row_count} data rows (within 40-70 range) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 - Expected ~55 rows, found {data_row_count}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Sum of Amount ~= 67500 (0.15 points)
    # The Electronics total in the pivot was 67500; the detail rows should sum to the same.
    try:
        amount_sum = 0.0
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4, values_only=True):
            val = row[0]
            if val is not None:
                try:
                    amount_sum += float(val)
                except (ValueError, TypeError):
                    pass
        # Allow 1% tolerance
        if abs(amount_sum - 67500) <= 675:
            print(f"PASS: Component 5 - Sum of Amount = {amount_sum:.2f} (~67500) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 - Sum of Amount = {amount_sum:.2f}, expected ~67500")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
