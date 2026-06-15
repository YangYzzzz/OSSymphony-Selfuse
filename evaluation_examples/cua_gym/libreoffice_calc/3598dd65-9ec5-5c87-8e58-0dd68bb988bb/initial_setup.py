"""
Initial Setup: Create sales records with a summary pivot-like table
Task ID: calc_pivot_047
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()

    # =========================================================
    # Sheet 1: SalesRecords — 200 rows of transaction data
    # =========================================================
    ws1 = wb.active
    ws1.title = 'SalesRecords'

    headers = ['ID', 'Category', 'Product', 'Amount']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Column widths
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 30
    ws1.column_dimensions['D'].width = 14

    # Define categories and products
    categories = {
        'Electronics': [
            'Wireless Bluetooth Headphones', 'USB-C Charging Cable', 'Portable Power Bank',
            'Smart Watch Band', 'Screen Protector Pack', 'Laptop Stand Adjustable',
            'Mechanical Keyboard', 'Wireless Mouse', 'HDMI Cable 6ft',
            'USB Flash Drive 64GB', 'Phone Case Premium', 'Webcam HD 1080p',
            'LED Desk Lamp', 'Tablet Stylus Pen', 'Car Phone Mount',
            'Bluetooth Speaker Mini', 'Lightning Adapter', 'Ethernet Cable 10ft',
            'Monitor Riser Stand', 'Cable Management Kit'
        ],
        'Clothing': [
            'Cotton Crew T-Shirt', 'Slim Fit Jeans', 'Running Sneakers',
            'Wool Blend Sweater', 'Waterproof Rain Jacket', 'Leather Belt Classic',
            'Athletic Shorts', 'Dress Socks Pack', 'Baseball Cap',
            'Winter Beanie', 'Polo Shirt', 'Cargo Pants'
        ],
        'Home & Garden': [
            'Stainless Steel Cookware Set', 'Memory Foam Pillow', 'Scented Candle Lavender',
            'Plant Pot Ceramic Large', 'Bath Towel Set Premium', 'Kitchen Knife Set',
            'Throw Blanket Fleece', 'Wall Clock Modern', 'Shower Curtain Linen',
            'Doormat Welcome', 'Picture Frame Set', 'Storage Bins Stackable'
        ],
        'Sports & Outdoors': [
            'Yoga Mat Non-Slip', 'Resistance Bands Set', 'Water Bottle Insulated',
            'Hiking Backpack 40L', 'Jump Rope Speed', 'Tennis Balls Pack',
            'Camping Lantern LED', 'Fishing Tackle Box', 'Bike Lock Heavy Duty',
            'Swim Goggles Anti-Fog'
        ],
        'Books & Media': [
            'Python Programming Guide', 'Mystery Novel Bestseller', 'Cookbook Mediterranean',
            'Journal Notebook Leather', 'Coloring Book Adults', 'History Atlas World',
            'Science Fiction Anthology', 'Business Strategy Manual', 'Art Sketchpad A4',
            'Language Learning Cards'
        ]
    }

    # We need exactly 55 Electronics rows summing to 67500
    # and 145 rows for other categories
    electronics_count = 55
    other_count = 200 - electronics_count  # 145

    # Generate Electronics amounts that sum to 67500
    electronics_amounts = []
    remaining = 67500.0
    for i in range(electronics_count - 1):
        # Average remaining per item
        avg = remaining / (electronics_count - i)
        amt = round(random.uniform(avg * 0.4, avg * 1.6), 2)
        amt = min(amt, remaining - (electronics_count - i - 1) * 50)  # leave room
        amt = max(amt, 50.0)
        electronics_amounts.append(amt)
        remaining -= amt
    electronics_amounts.append(round(remaining, 2))
    random.shuffle(electronics_amounts)

    # Generate other category rows
    other_cats = ['Clothing', 'Home & Garden', 'Sports & Outdoors', 'Books & Media']
    other_rows = []
    for i in range(other_count):
        cat = other_cats[i % len(other_cats)]
        product = random.choice(categories[cat])
        amount = round(random.uniform(15.0, 250.0), 2)
        other_rows.append((cat, product, amount))

    # Build all rows: interleave Electronics with others
    all_rows = []
    elec_idx = 0
    other_idx = 0

    for i in range(200):
        if elec_idx < electronics_count and (other_idx >= other_count or random.random() < 0.275):
            cat = 'Electronics'
            product = random.choice(categories['Electronics'])
            amount = electronics_amounts[elec_idx]
            elec_idx += 1
        elif other_idx < other_count:
            cat, product, amount = other_rows[other_idx]
            other_idx += 1
        else:
            cat = 'Electronics'
            product = random.choice(categories['Electronics'])
            amount = electronics_amounts[elec_idx]
            elec_idx += 1
        all_rows.append((cat, product, amount))

    # Write data rows
    for r, (cat, product, amount) in enumerate(all_rows, 2):
        ws1.cell(row=r, column=1, value=r - 1)  # ID = 1..200
        ws1.cell(row=r, column=2, value=cat)
        ws1.cell(row=r, column=3, value=product)
        cell_amt = ws1.cell(row=r, column=4, value=amount)
        cell_amt.number_format = '#,##0.00'

    # Auto-filter
    ws1.auto_filter.ref = "A1:D201"

    # =========================================================
    # Sheet 2: Summary — pivot-table-like summary
    # =========================================================
    ws2 = wb.create_sheet('Summary')

    # Title
    ws2.merge_cells('A1:B1')
    title_cell = ws2['A1']
    title_cell.value = 'Sales Summary by Category'
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    # Headers
    ws2['A3'] = 'Category'
    ws2['B3'] = 'Sum of Amount'
    ws2['A3'].font = Font(bold=True)
    ws2['B3'].font = Font(bold=True)
    ws2['A3'].fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    ws2['B3'].fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 18

    # Compute category totals from actual data
    cat_totals = {}
    for cat, product, amount in all_rows:
        cat_totals[cat] = cat_totals.get(cat, 0.0) + amount

    # Sort by category name for consistent display
    sorted_cats = sorted(cat_totals.keys())
    for i, cat in enumerate(sorted_cats):
        row = 4 + i
        ws2.cell(row=row, column=1, value=cat)
        amt_cell = ws2.cell(row=row, column=2, value=round(cat_totals[cat], 2))
        amt_cell.number_format = '#,##0.00'

    # Grand Total row
    grand_row = 4 + len(sorted_cats)
    ws2.cell(row=grand_row, column=1, value='Grand Total').font = Font(bold=True)
    grand_cell = ws2.cell(row=grand_row, column=2, value=round(sum(cat_totals.values()), 2))
    grand_cell.number_format = '#,##0.00'
    grand_cell.font = Font(bold=True)

    # Freeze panes
    ws1.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify Electronics total
    elec_total = sum(a for c, p, a in all_rows if c == 'Electronics')
    elec_count = sum(1 for c, p, a in all_rows if c == 'Electronics')
    print(f'Electronics: {elec_count} rows, total={elec_total:.2f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
