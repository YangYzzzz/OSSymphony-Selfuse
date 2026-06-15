"""
Reward Script: Federal income tax calculation with 2024 brackets
Task ID: calc_fin_tax_bracket_007
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.40): C2:C25 contain nested IF tax formulas with all 6 bracket boundaries
  - Component 2 (0.20): D2:D25 contain effective rate formulas (=Cn/Bn)
  - Component 3 (0.20): B2:B25 and C2:C25 are formatted as currency ($#,##0.00)
  - Component 4 (0.10): D2:D25 are formatted as percentage (0.0% or similar)
  - Component 5 (0.10): TaxBrackets sheet has a blue tab color set
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_tax_bracket_007'

# 2024 tax bracket boundary values that must appear in the IF formula
BRACKET_BOUNDARIES = ['23200', '94300', '201050', '383900', '487450']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition: Check required sheets exist ---
    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: 'Employees' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'TaxBrackets' not in wb.sheetnames:
        print("CRITICAL: 'TaxBrackets' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_emp = wb['Employees']
    ws_tax = wb['TaxBrackets']

    # -----------------------------------------------------------------------
    # Component 1: C2:C25 contain nested IF tax formulas (0.40 points)
    # Each formula must:
    #   - Start with =IF(
    #   - Reference all 5 bracket boundary values: 23200, 94300, 201050, 383900, 487450
    # -----------------------------------------------------------------------
    try:
        rows_with_if_formula = 0
        rows_with_all_boundaries = 0
        total_rows = 24  # rows 2..25

        for row in range(2, 26):
            cell = ws_emp.cell(row=row, column=3)
            val = cell.value

            # Must be a formula starting with =IF(
            if not isinstance(val, str) or not val.upper().startswith('=IF('):
                continue

            rows_with_if_formula += 1

            # Must reference all 5 bracket boundary values
            has_all_boundaries = all(boundary in val for boundary in BRACKET_BOUNDARIES)
            if has_all_boundaries:
                rows_with_all_boundaries += 1

        # Award points based on ratio of rows with correct formula
        if rows_with_all_boundaries == total_rows:
            print(f"PASS: Component 1 — All {total_rows} rows in C2:C25 have nested IF tax formulas "
                  f"with all bracket boundaries (0.40 pts)")
            total_score += 0.40
        elif rows_with_all_boundaries >= total_rows * 0.5:
            comp1_partial = round(0.40 * rows_with_all_boundaries / total_rows, 2)
            print(f"PARTIAL: Component 1 — {rows_with_all_boundaries}/{total_rows} rows have correct IF formulas "
                  f"(partial credit: {comp1_partial} pts)")
            if comp1_partial > 0:
                total_score += comp1_partial
        elif rows_with_if_formula > 0:
            print(f"PARTIAL: Component 1 — {rows_with_if_formula}/{total_rows} rows have =IF formulas "
                  f"but missing bracket boundaries ({rows_with_all_boundaries} complete)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — No nested IF tax formulas found in C2:C25 "
                  f"(expected =IF formulas with bracket values 23200, 94300, 201050, 383900, 487450)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: D2:D25 contain effective rate formulas =Cn/Bn (0.20 points)
    # Each formula must be of the form =Cn/Bn where n is the row number
    # -----------------------------------------------------------------------
    try:
        rows_with_rate_formula = 0
        total_rows = 24

        for row in range(2, 26):
            cell = ws_emp.cell(row=row, column=4)
            val = cell.value

            if not isinstance(val, str):
                continue

            # Check for =Cn/Bn pattern
            formula_upper = val.upper().strip()
            if formula_upper == f'=C{row}/B{row}':
                rows_with_rate_formula += 1
            elif re.match(r'^=C\d+/B\d+$', formula_upper):
                rows_with_rate_formula += 1

        if rows_with_rate_formula == total_rows:
            print(f"PASS: Component 2 — All {total_rows} rows in D2:D25 have effective rate formulas "
                  f"(=Cn/Bn) (0.20 pts)")
            total_score += 0.20
        elif rows_with_rate_formula >= total_rows * 0.5:
            comp2_partial = round(0.20 * rows_with_rate_formula / total_rows, 2)
            print(f"PARTIAL: Component 2 — {rows_with_rate_formula}/{total_rows} rows have rate formulas "
                  f"(partial credit: {comp2_partial} pts)")
            if comp2_partial > 0:
                total_score += comp2_partial
        else:
            print(f"FAIL: Component 2 — Only {rows_with_rate_formula}/{total_rows} rows have effective rate "
                  f"formulas in D2:D25 (expected =Cn/Bn pattern)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: B2:B25 and C2:C25 are formatted as currency $#,##0.00 (0.20 points)
    # Initial file has 'General' format for both B and C columns
    # -----------------------------------------------------------------------
    try:
        total_rows = 24
        b_currency_count = 0
        c_currency_count = 0

        for row in range(2, 26):
            b_fmt = ws_emp.cell(row=row, column=2).number_format
            c_fmt = ws_emp.cell(row=row, column=3).number_format

            # Check if format looks like currency (contains $ and # formatting)
            b_is_currency = '$' in b_fmt and '#' in b_fmt
            c_is_currency = '$' in c_fmt and '#' in c_fmt

            if b_is_currency:
                b_currency_count += 1
            if c_is_currency:
                c_currency_count += 1

        b_ok = b_currency_count == total_rows
        c_ok = c_currency_count == total_rows

        if b_ok and c_ok:
            print(f"PASS: Component 3 — B2:B25 and C2:C25 all formatted as currency (0.20 pts)")
            total_score += 0.20
        elif b_ok or c_ok:
            print(f"PARTIAL: Component 3 — Only one of B/C column has full currency format (0.10 pts)")
            total_score += 0.10
        elif (b_currency_count + c_currency_count) > 0:
            comp3_partial = round(0.20 * (b_currency_count + c_currency_count) / (2 * total_rows), 2)
            print(f"PARTIAL: Component 3 — B currency: {b_currency_count}/{total_rows}, "
                  f"C currency: {c_currency_count}/{total_rows} "
                  f"(partial credit: {comp3_partial} pts)")
            if comp3_partial > 0:
                total_score += comp3_partial
        else:
            print(f"FAIL: Component 3 — No currency formatting found in B2:B25 or C2:C25 "
                  f"(expected $#,##0.00 format)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: D2:D25 are formatted as percentage (0.10 points)
    # Initial file has 'General' format for D column
    # -----------------------------------------------------------------------
    try:
        total_rows = 24
        d_pct_count = 0

        for row in range(2, 26):
            d_fmt = ws_emp.cell(row=row, column=4).number_format
            # Accept any percentage format: 0%, 0.0%, 0.00%, etc.
            if '%' in d_fmt:
                d_pct_count += 1

        if d_pct_count == total_rows:
            print(f"PASS: Component 4 — All {total_rows} rows in D2:D25 formatted as percentage (0.10 pts)")
            total_score += 0.10
        elif d_pct_count > 0:
            comp4_partial = round(0.10 * d_pct_count / total_rows, 2)
            print(f"PARTIAL: Component 4 — {d_pct_count}/{total_rows} D cells have percentage format "
                  f"(partial credit: {comp4_partial} pts)")
            if comp4_partial > 0:
                total_score += comp4_partial
        else:
            print(f"FAIL: Component 4 — No percentage formatting in D2:D25 "
                  f"(expected 0.0% or similar format)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: TaxBrackets sheet tab has a blue color set (0.10 points)
    # Initial file has no tab color (None)
    # Golden file has tab color rgb='004472C4' (blue: R=44, G=72, B=C4)
    # -----------------------------------------------------------------------
    try:
        tab_color = ws_tax.sheet_properties.tabColor

        if tab_color is not None:
            rgb_val = tab_color.rgb if tab_color.rgb else ''
            print(f"PASS: Component 5 — TaxBrackets tab color is set "
                  f"(rgb={rgb_val}, 0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — TaxBrackets sheet has no tab color set "
                  f"(expected a blue tab color)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
