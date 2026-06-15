"""
Initial Setup: Federal income tax calculation setup with employee salaries and tax bracket table
Task ID: calc_fin_tax_bracket_007
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_tax_bracket_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Employees ---
    ws_emp = wb.active
    ws_emp.title = 'Employees'

    # Headers
    headers = ['Employee', 'Annual Salary', 'Federal Tax', 'Effective Rate']
    for col, h in enumerate(headers, 1):
        ws_emp.cell(row=1, column=col, value=h)
        ws_emp.cell(row=1, column=col).font = Font(bold=True)

    # Realistic employee data with varied salaries across all tax brackets
    employees = [
        ('Sarah Chen', 45000),
        ('Marcus Johnson', 128500),
        ('Emily Rodriguez', 22000),
        ('David Kim', 195000),
        ('Jennifer Walsh', 75000),
        ('Robert Nguyen', 310000),
        ('Amanda Foster', 58000),
        ('Christopher Lee', 420000),
        ('Michelle Torres', 33500),
        ('Brandon Mitchell', 165000),
        ('Ashley Peterson', 88000),
        ('Kevin Thompson', 520000),
        ('Stephanie Brown', 47500),
        ('Daniel Rivera', 102000),
        ('Laura Martinez', 19500),
        ('James Wilson', 245000),
        ('Rachel Anderson', 67000),
        ('Matthew Clark', 370000),
        ('Nicole Harris', 55000),
        ('Andrew White', 138000),
        ('Megan Lewis', 92000),
        ('Ryan Jackson', 480000),
        ('Brittany Young', 31000),
        ('Tyler Scott', 175000),
    ]

    for r, (name, salary) in enumerate(employees, 2):
        ws_emp.cell(row=r, column=1, value=name)
        ws_emp.cell(row=r, column=2, value=salary)
        # C and D columns intentionally left empty (task is to add formulas)

    # Set column widths for readability
    ws_emp.column_dimensions['A'].width = 22
    ws_emp.column_dimensions['B'].width = 16
    ws_emp.column_dimensions['C'].width = 16
    ws_emp.column_dimensions['D'].width = 16

    # --- Sheet 2: TaxBrackets ---
    ws_tax = wb.create_sheet('TaxBrackets')

    # Headers for bracket table
    bracket_headers = ['Min Income', 'Max Income', 'Rate']
    for col, h in enumerate(bracket_headers, 1):
        ws_tax.cell(row=1, column=col, value=h)
        ws_tax.cell(row=1, column=col).font = Font(bold=True)

    # 2024 MFJ (Married Filing Jointly) tax brackets
    brackets = [
        (0, 23200, 0.10),
        (23200, 94300, 0.12),
        (94300, 201050, 0.22),
        (201050, 383900, 0.24),
        (383900, 487450, 0.32),
        (487450, 9999999, 0.37),
    ]

    for r, (min_inc, max_inc, rate) in enumerate(brackets, 2):
        ws_tax.cell(row=r, column=1, value=min_inc)
        ws_tax.cell(row=r, column=2, value=max_inc)
        ws_tax.cell(row=r, column=3, value=rate)

    # Format Rate column as percentage
    for r in range(2, 8):
        ws_tax.cell(row=r, column=3).number_format = '0%'

    # Set column widths
    ws_tax.column_dimensions['A'].width = 16
    ws_tax.column_dimensions['B'].width = 16
    ws_tax.column_dimensions['C'].width = 10

    # Note: TaxBrackets tab color is NOT set here — that's the task to complete

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet Employees: {len(employees)} employees, C and D columns empty')
    print(f'  Sheet TaxBrackets: 2024 MFJ brackets in A1:C7')
    print(f'  TaxBrackets tab color: none (to be set blue by agent)')


create_initial()
