"""
Initial Setup: Rename named range 'data_2023' to 'data_2024' and update reference
Task ID: calc_nrv_022
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Monthly Revenue Data ---
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    ws.cell(row=1, column=1, value='Month')
    ws.cell(row=1, column=2, value='Revenue ($)')

    # Generate 59 rows of realistic monthly/weekly revenue data (B2:B60)
    import random
    random.seed(42)

    labels = []
    # Create date-based labels for 59 periods (Jan 2019 through Nov 2023)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    year = 2019
    month_idx = 0
    for i in range(59):
        labels.append(f'{months[month_idx]} {year}')
        month_idx += 1
        if month_idx >= 12:
            month_idx = 0
            year += 1

    base_revenue = 42000
    for i, label in enumerate(labels):
        row = i + 2
        ws.cell(row=row, column=1, value=label)
        # Slightly trending upward with noise
        revenue = round(base_revenue + i * 350 + random.uniform(-3000, 3000), 2)
        ws.cell(row=row, column=2, value=revenue)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18

    # Bold header row
    from openpyxl.styles import Font
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # --- Define named range 'data_2023' referring to Sheet1.$B$2:$B$50 ---
    defn = DefinedName('data_2023', attr_text="Sheet1!$B$2:$B$50")
    wb.defined_names.add(defn)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify named range was saved
    for name in wb.defined_names:
        dn = wb.defined_names[name]
        print(f'  Named range: {dn.name} = {dn.attr_text}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
