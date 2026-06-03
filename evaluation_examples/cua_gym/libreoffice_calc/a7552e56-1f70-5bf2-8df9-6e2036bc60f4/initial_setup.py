"""
Initial Setup: Create spreadsheet with 'Formulas' sheet (unprotected) for calc_adv_protect_sheet_named_050
Task ID: calc_adv_protect_sheet_named_050
Domain: libreoffice_calc

The 'Formulas' sheet must:
- Have Parameter names in column A
- Have Input values in column B (B2:B10 unlocked, all others default locked)
- Have Calculated results (formulas) in column C
- Have Explanation notes in column D
- Be UNPROTECTED (no sheet protection)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_protect_sheet_named_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Formulas ---
    ws = wb.active
    ws.title = 'Formulas'

    # --- Headers ---
    headers = ['Parameter', 'Input Value', 'Calculated Result', 'Explanation']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9EAF7', end_color='FFD9EAF7', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Data rows ---
    # Column A: Parameter names
    # Column B: Input values (will be set unlocked=False for B2:B10)
    # Column C: Calculated results using formulas referencing column B
    # Column D: Explanation notes

    parameters = [
        'Annual Revenue',
        'Tax Rate (%)',
        'Operating Expenses',
        'Depreciation',
        'Interest Expense',
        'Other Income',
        'Discount Rate (%)',
        'Growth Rate (%)',
        'Inflation Rate (%)',
    ]

    input_values = [
        2500000,
        28,
        850000,
        120000,
        45000,
        32000,
        8.5,
        5.2,
        3.1,
    ]

    explanations = [
        'Total annual revenue before deductions',
        'Effective corporate tax rate applied to taxable income',
        'Total operating costs excluding depreciation',
        'Annual depreciation of fixed assets',
        'Annual interest on outstanding debt',
        'Additional income from non-core activities',
        'Used to calculate net present value of future cash flows',
        'Expected year-over-year revenue growth',
        'Consumer price index annual growth rate',
    ]

    for i in range(9):
        row = i + 2
        param = parameters[i]
        inp_val = input_values[i]
        expl = explanations[i]

        # Column A: Parameter name (locked by default)
        ws.cell(row=row, column=1, value=param)

        # Column B: Input value — set locked=False as per context
        b_cell = ws.cell(row=row, column=2, value=inp_val)
        b_cell.protection = Protection(locked=False)

        # Column C: Formula referencing B column
        if i == 0:  # Annual Revenue — show raw (no formula)
            ws.cell(row=row, column=3, value=f'=B{row}')
        elif i == 1:  # Tax Rate
            # Net after tax = B2 * (1 - B3/100) conceptually; show tax amount
            ws.cell(row=row, column=3, value=f'=B2*B{row}/100')
        elif i == 2:  # Operating Expenses
            ws.cell(row=row, column=3, value=f'=B{row}')
        elif i == 3:  # Depreciation
            ws.cell(row=row, column=3, value=f'=B{row}')
        elif i == 4:  # Interest Expense
            ws.cell(row=row, column=3, value=f'=B{row}')
        elif i == 5:  # Other Income
            ws.cell(row=row, column=3, value=f'=B{row}')
        elif i == 6:  # Discount Rate — NPV factor
            ws.cell(row=row, column=3, value=f'=1/(1+B{row}/100)')
        elif i == 7:  # Growth Rate
            ws.cell(row=row, column=3, value=f'=B2*(1+B{row}/100)')
        elif i == 8:  # Inflation Rate
            ws.cell(row=row, column=3, value=f'=B2/(1+B{row}/100)')

        # Column D: Explanation
        ws.cell(row=row, column=4, value=expl)

    # Summary rows
    ws.cell(row=12, column=1, value='EBITDA')
    ws.cell(row=12, column=3, value='=B2-B4-B5')
    ws.cell(row=12, column=4, value='Earnings Before Interest, Taxes, Depreciation & Amortization')

    ws.cell(row=13, column=1, value='Net Income')
    ws.cell(row=13, column=3, value='=B2-B4-B5-B6-(B2-B4-B5-B6)*B3/100')
    ws.cell(row=13, column=4, value='Revenue minus all expenses and taxes')

    ws.cell(row=14, column=1, value='Profit Margin (%)')
    ws.cell(row=14, column=3, value='=C13/B2*100')
    ws.cell(row=14, column=4, value='Net income as percentage of revenue')

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 45

    # Row height for header
    ws.row_dimensions[1].height = 22

    # Sheet is NOT protected (that's the task)
    # ws.protection is default (sheet=False = no protection)

    # Add a second sheet for data context
    ws2 = wb.create_sheet('Reference')
    ws2['A1'] = 'Fiscal Year'
    ws2['B1'] = 'Period'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['A2'] = 2025
    ws2['B2'] = 'Annual'
    ws2['A3'] = 'Currency'
    ws2['B3'] = 'USD'
    ws2['A4'] = 'Last Updated'
    ws2['B4'] = '2025-01-15'
    ws2['A5'] = 'Prepared By'
    ws2['B5'] = 'Finance Department'
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet names: {wb.sheetnames}')
    print(f'Formulas sheet protection enabled: {ws.protection.sheet}')
    print(f'B2:B10 locked=False (unlocked for editing)')


create_initial()
