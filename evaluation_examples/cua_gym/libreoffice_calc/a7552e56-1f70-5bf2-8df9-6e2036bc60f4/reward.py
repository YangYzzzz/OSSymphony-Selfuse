"""
Reward Script: Protect the 'Formulas' sheet with password 'NoEdit!' allowing navigation
Task ID: calc_adv_protect_sheet_named_050
Domain: libreoffice_calc

Task: Protect the 'Formulas' sheet with password 'NoEdit!' but allow users to select
both locked and unlocked cells so they can still navigate and read the data.

Scoring:
- Component 1 (0.4): Sheet 'Formulas' is protected with the correct password 'NoEdit!'
- Component 2 (0.3): Protection allows selecting both locked and unlocked cells (navigation enabled)
- Component 3 (0.3): B2:B10 cells remain unlocked; other cells (A col, C col, D col, B1) remain locked
"""

import os
import openpyxl
import openpyxl.worksheet.protection as prot_module

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_protect_sheet_named_050'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Formulas' sheet must exist
    if 'Formulas' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Formulas' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Formulas']

    # Component 1: Sheet 'Formulas' is protected with password 'NoEdit!' (0.4 points)
    # This FAILS on initial (protection.sheet=False) → PASSES on golden (protection.sheet=True, _password='CBA4')
    try:
        is_protected = ws.protection.sheet
        stored_password = ws.protection._password

        # Compute the expected legacy hash for 'NoEdit!'
        # openpyxl uses a legacy XOR-based hash for XLSX sheet passwords
        p_check = prot_module.SheetProtection()
        p_check.set_password('NoEdit!')
        expected_hash = p_check._password  # Should be 'CBA4'

        if is_protected and stored_password == expected_hash:
            print(f"PASS: Component 1 — Sheet 'Formulas' is protected with password 'NoEdit!' "
                  f"(hash={stored_password}) (0.4 pts)")
            total_score += 0.4
        elif is_protected and stored_password != expected_hash:
            print(f"FAIL: Component 1 — Sheet is protected but password hash is wrong. "
                  f"Expected hash '{expected_hash}', found '{stored_password}'")
        else:
            print(f"FAIL: Component 1 — Sheet 'Formulas' is NOT protected "
                  f"(protection.sheet={is_protected})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Protection allows selecting both locked and unlocked cells (0.3 points)
    # In openpyxl: selectLockedCells=False means users ARE allowed to select locked cells
    #              selectUnlockedCells=False means users ARE allowed to select unlocked cells
    # So BOTH must be False (i.e., neither restriction is active → both types selectable)
    # This FAILS on initial (no protection at all) → PASSES on golden
    try:
        # Only meaningful to check if sheet is protected
        if not ws.protection.sheet:
            print(f"FAIL: Component 2 — Sheet is not protected, cannot verify selection options")
        else:
            select_locked = ws.protection.selectLockedCells
            select_unlocked = ws.protection.selectUnlockedCells

            # selectLockedCells=False means "do NOT restrict selecting locked cells" → users CAN select locked
            # selectUnlockedCells=False means "do NOT restrict selecting unlocked cells" → users CAN select unlocked
            allow_locked = (select_locked == False)
            allow_unlocked = (select_unlocked == False)

            if allow_locked and allow_unlocked:
                print(f"PASS: Component 2 — Protection allows selecting both locked cells "
                      f"(selectLockedCells={select_locked}) and unlocked cells "
                      f"(selectUnlockedCells={select_unlocked}) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Selection not fully enabled. "
                      f"selectLockedCells={select_locked} (want False), "
                      f"selectUnlockedCells={select_unlocked} (want False)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B2:B10 are unlocked; other cells (A col, C col, D col, B1) remain locked (0.3 points)
    # B2:B10 should have locked=False (pre-unlocked input cells)
    # Header and formula cells should have locked=True (default)
    # This FAILS on initial (no protection context, but more importantly B2:B10 being unlocked
    # is meaningless without sheet protection — the reward verifies cell-level lock settings
    # which are already set in initial). Wait — let me re-check: in the initial file B2:B10 are
    # also locked=False. So this cannot be a differentiating component on its own.
    #
    # CORRECTION: Per the "Only Score Task-Introduced Changes" rule, if B2:B10 lock status
    # is ALREADY False in the initial file, checking that alone would score points on initial too.
    # We must combine this with the sheet-protection check.
    # Component 3 checks: sheet IS protected AND B2:B10 are unlocked AND locked cells are locked.
    # This compound check FAILS on initial (sheet not protected) and PASSES on golden.
    try:
        if not ws.protection.sheet:
            print(f"FAIL: Component 3 — Sheet is not protected; cell lock verification skipped")
        else:
            # Check B2:B10 are unlocked
            b_cells_unlocked = True
            for row in range(2, 11):
                cell = ws.cell(row=row, column=2)
                if cell.protection.locked is not False:
                    b_cells_unlocked = False
                    print(f"  FAIL: B{row} should be unlocked but locked={cell.protection.locked}")
                    break

            # Check that other cells are locked (A1, B1, C2, D2 as representative samples)
            locked_cells_locked = True
            check_coords = [('A', 1), ('B', 1), ('C', 2), ('D', 2), ('A', 2)]
            for col_letter, row_num in check_coords:
                from openpyxl.utils import column_index_from_string
                col_idx = column_index_from_string(col_letter)
                cell = ws.cell(row=row_num, column=col_idx)
                # Default locked is True; if explicitly set to False it's a problem
                cell_locked = cell.protection.locked
                # locked=None means default (True), locked=True means explicitly True
                if cell_locked is False:
                    locked_cells_locked = False
                    print(f"  FAIL: {col_letter}{row_num} should be locked but locked=False")
                    break

            if b_cells_unlocked and locked_cells_locked:
                print(f"PASS: Component 3 — Sheet is protected with correct cell-level locks: "
                      f"B2:B10 unlocked, header/formula/note cells locked (0.3 pts)")
                total_score += 0.3
            elif not b_cells_unlocked:
                print(f"FAIL: Component 3 — B2:B10 cells should be unlocked but some are locked")
            else:
                print(f"FAIL: Component 3 — Some cells that should be locked are unlocked")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
