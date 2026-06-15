"""
Reward Script: Replace VLOOKUP formulas with static values in column E
Task ID: calc_tbl_039
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): No formulas remain in E2:E500
  Component 2 (0.3): Values match LookupSheet data (spot-check)
  Component 3 (0.2): All 499 cells in E2:E500 are non-None numeric values
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_039'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that E2:E500 formulas have been replaced with static values.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data'] if 'Data' in wb.sheetnames else wb.active

    # Component 1: No formulas remain in E2:E500 (0.5 points)
    # This is the core task requirement - formulas must be replaced
    try:
        formula_count = 0
        for r in range(2, 501):
            val = ws.cell(row=r, column=5).value
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1

        if formula_count == 0:
            print(f"PASS: Component 1 — No formulas in E2:E500 (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Found {formula_count} formulas still in E2:E500")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Values match LookupSheet data (0.3 points)
    # Verify the paste-as-values produced correct results by cross-checking
    # against LookupSheet column C (Salary)
    try:
        if 'LookupSheet' not in wb.sheetnames:
            print("FAIL: Component 2 — LookupSheet not found, cannot cross-check")
        else:
            ls = wb['LookupSheet']
            # Build lookup dictionary: Employee ID -> Salary (col C)
            lookup = {}
            for r in range(2, ls.max_row + 1):
                key = ls.cell(row=r, column=1).value
                sal = ls.cell(row=r, column=3).value
                if key is not None and sal is not None:
                    lookup[key] = float(sal)

            # Spot-check a sample of rows for correctness
            matches = 0
            checked = 0
            for r in range(2, 501):
                emp_id = ws.cell(row=r, column=1).value
                e_val = ws.cell(row=r, column=5).value
                expected = lookup.get(emp_id)
                if e_val is not None and expected is not None:
                    checked += 1
                    try:
                        if abs(float(e_val) - expected) < 0.01:
                            matches += 1
                    except (ValueError, TypeError):
                        pass

            if checked > 0 and matches == checked:
                print(f"PASS: Component 2 — All {matches}/{checked} values match LookupSheet (0.3 pts)")
                total_score += 0.3
            elif checked > 0:
                ratio = matches / checked
                partial = round(0.3 * ratio, 2)
                print(f"PARTIAL: Component 2 — {matches}/{checked} values match ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Could not verify any values")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 499 cells in E2:E500 have non-None numeric values (0.2 points)
    # Ensures no data was lost during the conversion
    try:
        numeric_count = 0
        none_count = 0
        for r in range(2, 501):
            val = ws.cell(row=r, column=5).value
            if val is None:
                none_count += 1
            elif isinstance(val, (int, float)):
                numeric_count += 1

        if numeric_count == 499:
            print(f"PASS: Component 3 — All 499 cells have numeric values (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — numeric={numeric_count}, none={none_count}, expected 499 numeric")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
