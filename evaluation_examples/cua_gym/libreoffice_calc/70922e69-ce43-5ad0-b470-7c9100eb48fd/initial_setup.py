"""
Initial Setup: Replace VLOOKUP formulas with static values for performance
Task ID: calc_tbl_039
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# --- Lookup data ---
# 500 employees with IDs, names, salaries, locations
first_names = [
    "Sarah", "Marcus", "Priya", "James", "Mei", "Carlos", "Fatima", "David",
    "Yuki", "Elena", "Kwame", "Amara", "Lucas", "Zara", "Thomas", "Nia",
    "Raj", "Sofia", "Liam", "Aaliyah", "Viktor", "Chen", "Olivia", "Hassan",
    "Maria", "Derek", "Anika", "Patrick", "Isla", "Mohammed", "Grace", "Andrei",
    "Keiko", "Daniel", "Nadia", "Samuel", "Luz", "Erik", "Chioma", "Anton",
    "Rosa", "Kai", "Ingrid", "Felix", "Adaeze", "Martin", "Suki", "Oscar",
    "Leila", "Bjorn"
]
last_names = [
    "Chen", "Johnson", "Patel", "Williams", "Tanaka", "Rodriguez", "Al-Hassan",
    "Kim", "Nakamura", "Petrova", "Mensah", "Okafor", "Silva", "Novak",
    "Thompson", "Diallo", "Gupta", "Martinez", "O'Brien", "Washington",
    "Ivanov", "Wei", "Brown", "Ahmed", "Garcia", "Campbell", "Mueller",
    "Johansson", "Singh", "Kowalski", "Lopez", "Park", "Svensson", "Moreau",
    "Taniguchi", "Fischer", "Osei", "Fernandez", "Larsson", "Nguyen"
]
departments = [
    "Engineering", "Marketing", "Finance", "Operations", "Human Resources",
    "Sales", "Product", "Legal", "Customer Support", "Research"
]
locations = [
    "New York", "San Francisco", "London", "Tokyo", "Berlin",
    "Singapore", "Toronto", "Sydney", "Mumbai", "Sao Paulo",
    "Chicago", "Seattle", "Austin", "Dublin", "Amsterdam"
]

random.seed(42)

# Build lookup entries for 500 employees
employee_ids = [f"EMP-{i:04d}" for i in range(1, 501)]
employee_names = [f"{random.choice(first_names)} {random.choice(last_names)}" for _ in range(500)]
employee_salaries = [round(random.uniform(42000, 185000), 2) for _ in range(500)]
employee_locations = [random.choice(locations) for _ in range(500)]
employee_departments = [random.choice(departments) for _ in range(500)]
hire_dates = []
for _ in range(500):
    year = random.randint(2018, 2025)
    month = random.randint(1, 12)
    day = random.randint(1, 28)
    hire_dates.append(f"{year}-{month:02d}-{day:02d}")


def create_initial():
    wb = openpyxl.Workbook()

    # --- LookupSheet ---
    ws_lookup = wb.active
    ws_lookup.title = "LookupSheet"

    # Headers
    lookup_headers = ["Employee ID", "Name", "Salary", "Location"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(lookup_headers, 1):
        cell = ws_lookup.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    for i in range(500):
        ws_lookup.cell(row=i + 2, column=1, value=employee_ids[i])
        ws_lookup.cell(row=i + 2, column=2, value=employee_names[i])
        ws_lookup.cell(row=i + 2, column=3, value=employee_salaries[i])
        ws_lookup.cell(row=i + 2, column=3).number_format = '$#,##0.00'
        ws_lookup.cell(row=i + 2, column=4, value=employee_locations[i])

    ws_lookup.column_dimensions["A"].width = 14
    ws_lookup.column_dimensions["B"].width = 22
    ws_lookup.column_dimensions["C"].width = 14
    ws_lookup.column_dimensions["D"].width = 16

    # --- Data Sheet (main sheet with VLOOKUPs) ---
    ws_data = wb.create_sheet("Data", 0)  # insert at position 0 so it's the first/active sheet

    data_headers = ["Employee ID", "Department", "Hire Date", "Status", "Annual Salary"]
    for col, h in enumerate(data_headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    statuses = ["Active", "Active", "Active", "Active", "Active",
                "Active", "Active", "On Leave", "Active", "Active"]

    for i in range(500):  # rows 2..501
        row = i + 2
        ws_data.cell(row=row, column=1, value=employee_ids[i])
        ws_data.cell(row=row, column=2, value=employee_departments[i])
        ws_data.cell(row=row, column=3, value=hire_dates[i])
        ws_data.cell(row=row, column=4, value=statuses[i % len(statuses)])
        # Column E: VLOOKUP formula referencing LookupSheet
        ws_data.cell(row=row, column=5,
                     value=f"=VLOOKUP(A{row},LookupSheet.A:D,3,0)")
        ws_data.cell(row=row, column=5).number_format = '$#,##0.00'

    ws_data.column_dimensions["A"].width = 14
    ws_data.column_dimensions["B"].width = 18
    ws_data.column_dimensions["C"].width = 14
    ws_data.column_dimensions["D"].width = 12
    ws_data.column_dimensions["E"].width = 16

    ws_data.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
