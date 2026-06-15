"""
Initial Setup: Create a trend report spreadsheet with numeric trend indicators
Task ID: calc_gfl_063
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trends'

    # --- Headers ---
    headers = ['Metric', 'Previous Period', 'Current Period', 'Change', 'Trend']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data: 29 metrics (rows 2-30) ---
    metrics_data = [
        ('Website Traffic', 145200, 162800, 17600, 12.1),
        ('Bounce Rate (%)', 42.3, 38.7, -3.6, -8.5),
        ('Avg Session Duration (min)', 3.8, 4.2, 0.4, 10.5),
        ('New User Signups', 2340, 2890, 550, 23.5),
        ('Customer Retention Rate (%)', 87.5, 87.5, 0.0, 0.0),
        ('Revenue per Visit ($)', 3.45, 3.72, 0.27, 7.8),
        ('Cart Abandonment Rate (%)', 68.2, 71.5, 3.3, 4.8),
        ('Email Open Rate (%)', 22.1, 19.8, -2.3, -10.4),
        ('Social Media Followers', 34500, 36200, 1700, 4.9),
        ('Ad Click-Through Rate (%)', 1.8, 1.8, 0.0, 0.0),
        ('Customer Support Tickets', 892, 756, -136, -15.2),
        ('Average Order Value ($)', 67.30, 72.15, 4.85, 7.2),
        ('Page Load Time (sec)', 2.4, 2.1, -0.3, -12.5),
        ('Mobile Conversion Rate (%)', 2.1, 2.8, 0.7, 33.3),
        ('Inventory Turnover', 5.2, 5.2, 0.0, 0.0),
        ('Product Return Rate (%)', 8.7, 9.1, 0.4, 4.6),
        ('Customer Lifetime Value ($)', 342.50, 378.20, 35.70, 10.4),
        ('Net Promoter Score', 45, 52, 7, 15.6),
        ('Organic Search Traffic', 58400, 54200, -4200, -7.2),
        ('Paid Search ROI (%)', 285, 310, 25, 8.8),
        ('Email Subscriber Growth', 1250, 1250, 0, 0.0),
        ('Repeat Purchase Rate (%)', 31.4, 35.8, 4.4, 14.0),
        ('Warehouse Fulfillment Time (hrs)', 18.5, 16.2, -2.3, -12.4),
        ('Supplier Lead Time (days)', 14, 14, 0, 0.0),
        ('Gross Margin (%)', 42.8, 44.1, 1.3, 3.0),
        ('Employee Productivity Index', 78.3, 81.7, 3.4, 4.3),
        ('Quality Defect Rate (%)', 1.2, 0.9, -0.3, -25.0),
        ('Market Share (%)', 12.4, 12.4, 0.0, 0.0),
        ('Customer Acquisition Cost ($)', 45.20, 41.80, -3.40, -7.5),
    ]

    for r, (metric, prev, curr, change, trend) in enumerate(metrics_data, 2):
        ws.cell(row=r, column=1, value=metric)
        ws.cell(row=r, column=2, value=prev)
        ws.cell(row=r, column=3, value=curr)
        ws.cell(row=r, column=4, value=change)
        ws.cell(row=r, column=5, value=trend)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    # --- Number formats for data cells ---
    for r in range(2, 31):
        ws.cell(row=r, column=2).number_format = '#,##0.00'
        ws.cell(row=r, column=3).number_format = '#,##0.00'
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5).number_format = '0.0'

    # --- NO conditional formatting in initial state ---

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
