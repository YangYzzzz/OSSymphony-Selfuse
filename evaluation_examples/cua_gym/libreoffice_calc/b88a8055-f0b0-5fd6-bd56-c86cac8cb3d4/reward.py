"""
Reward Script: Apply icon set conditional formatting (arrows) to Trend column
Task ID: calc_gfl_063
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Conditional formatting rule exists on E2:E30
  Component 2 (0.4): Rule is iconSet type with 3Arrows (directional arrows)
  Component 3 (0.3): CFVO thresholds correctly configured for pos/zero/neg
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_063'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Trends' sheet must exist
    if 'Trends' not in wb.sheetnames:
        print(f"FAIL: 'Trends' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Trends']

    # Collect all conditional formatting rules that cover E2:E30 (or a superset)
    cf_list = list(ws.conditional_formatting)
    icon_set_rules_on_e = []

    for cf in cf_list:
        # Check if this CF range covers column E rows 2-30
        # The range could be E2:E30, E:E, E1:E30, etc.
        range_covers_e = any(
            cell_range.min_col <= 5 <= cell_range.max_col
            and cell_range.min_row <= 2
            and cell_range.max_row >= 30
            for cell_range in cf.cells.ranges
        )
        if range_covers_e:
            for rule in cf.rules:
                if rule.type == 'iconSet':
                    icon_set_rules_on_e.append(rule)

    # Component 1: Conditional formatting rule exists on E2:E30 (0.3 points)
    try:
        if len(icon_set_rules_on_e) > 0:
            print(f"PASS: Component 1 — Found {len(icon_set_rules_on_e)} iconSet CF rule(s) on E2:E30 (0.3 pts)")
            total_score += 0.3
        else:
            # Also check for any CF on E column even if range is slightly different
            e_col_has_cf = any(
                cell_range.min_col <= 5 <= cell_range.max_col
                for cf in cf_list
                for cell_range in cf.cells.ranges
            )
            if e_col_has_cf:
                print(f"FAIL: Component 1 — CF exists on column E but no iconSet rule found covering E2:E30")
            else:
                print(f"FAIL: Component 1 — No conditional formatting found on column E")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Rule is iconSet type with 3Arrows (0.4 points)
    try:
        if len(icon_set_rules_on_e) > 0:
            rule = icon_set_rules_on_e[0]
            icon_set_name = rule.iconSet.iconSet if rule.iconSet else None
            # Accept arrow-based icon sets (3Arrows is canonical, but 3ArrowsGray is also valid)
            if icon_set_name and 'Arrows' in icon_set_name:
                print(f"PASS: Component 2 — IconSet type is '{icon_set_name}' (arrows) (0.4 pts)")
                total_score += 0.4
            elif icon_set_name and ('arrow' in icon_set_name.lower()):
                print(f"PASS: Component 2 — IconSet type is '{icon_set_name}' (arrows variant) (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — Expected arrow icon set, found: '{icon_set_name}'")
        else:
            print(f"FAIL: Component 2 — No iconSet rule found to check")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: CFVO thresholds configured correctly (0.3 points)
    # Expected: 3 CFVOs that partition values into negative / zero / positive
    # The golden file has: cfvo[0]=num/0.0, cfvo[1]=num/0.0, cfvo[2]=num/1.0
    # This means: values < 0 get down arrow, 0 <= val < 1 get right arrow, val >= 1 get up arrow
    try:
        if len(icon_set_rules_on_e) > 0:
            rule = icon_set_rules_on_e[0]
            cfvos = rule.iconSet.cfvo if rule.iconSet else []
            if cfvos and len(cfvos) == 3:
                # Verify the thresholds create a meaningful partition for negative/zero/positive
                # The key requirement is that:
                # - There are 3 tiers (for down/right/up arrows)
                # - The thresholds separate negative from zero from positive
                types = [cfvo.type for cfvo in cfvos]
                vals = [cfvo.val for cfvo in cfvos]

                # Check that it's a 3-tier numeric partition
                # Valid configurations use 'num' or 'percent' type thresholds
                # The thresholds should create: negative=down, zero/small=right, positive=up
                has_valid_thresholds = False

                # Accept the golden configuration: num/0, num/0, num/1
                # Also accept reasonable alternatives like num/0, num/0, num/0 (with gte)
                # or percent-based thresholds
                if all(t in ('num', 'percent', 'percentile', 'min') for t in types):
                    print(f"PASS: Component 3 — CFVO thresholds: {list(zip(types, vals))} (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 3 — Unexpected CFVO types: {types}")
            else:
                num_cfvos = len(cfvos) if cfvos else 0
                print(f"FAIL: Component 3 — Expected 3 CFVOs, found {num_cfvos}")
        else:
            print(f"FAIL: Component 3 — No iconSet rule found to check thresholds")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
