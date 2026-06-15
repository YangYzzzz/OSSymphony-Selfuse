"""
Initial Setup: Tokyo sushi restaurant lookup task
Task ID: osworld_multi_apps_restaurant_lookup_002
Domain: libreoffice_calc (multi-app: gedit + LibreOffice Calc)

Creates:
  - /home/user/Desktop/tokyo_sushi.txt  (list of 5 restaurants, opened in gedit)
  - /home/user/Desktop/SUSHI_SPOTS.xlsx (spreadsheet with headers + restaurant names,
                                          columns B/C/D empty, opened in LibreOffice Calc)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_002'
TXT_FILE = f'{WORKDIR}/tokyo_sushi.txt'
XLSX_FILE = f'{WORKDIR}/SUSHI_SPOTS.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(WORKDIR, exist_ok=True)

    # --- Create tokyo_sushi.txt ---
    restaurants = [
        "Harutaka",
        "Sukiyabashi Jiro Honten",
        "Sushi Saito",
        "Sushi Yoshitake",
        "Sushi Sho",
    ]
    with open(TXT_FILE, 'w', encoding='utf-8') as f:
        f.write("Famous Tokyo Sushi Restaurants to Visit\n")
        f.write("=" * 40 + "\n\n")
        for i, name in enumerate(restaurants, 1):
            f.write(f"{i}. {name}\n")
    print(f'Text file created: {TXT_FILE}')

    # --- Create SUSHI_SPOTS.xlsx ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sushi Spots"

    # Headers in row 1
    headers = ['Name', 'Address', 'Website', 'Phone']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Restaurant names pre-filled in column A (rows 2-6), B/C/D left empty
    for row_idx, name in enumerate(restaurants, 2):
        ws.cell(row=row_idx, column=1, value=name)
        # Columns B (Address), C (Website), D (Phone) intentionally left empty

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18

    wb.save(XLSX_FILE)
    print(f'Spreadsheet created: {XLSX_FILE}')

    # --- GUI-ready startup ---
    # Open tokyo_sushi.txt in gedit
    launch_gui(f'gedit "{TXT_FILE}"', delay_sec=2.0)

    # Open SUSHI_SPOTS.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{XLSX_FILE}"', delay_sec=3.0)

    print('GUI_READY: launched gedit and LibreOffice Calc with DISPLAY=:0')


create_initial()
