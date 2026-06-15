"""
Reward Script: Tokyo Sushi Restaurants Lookup
Task ID: osworld_multi_apps_restaurant_lookup_002
Domain: libreoffice_calc
Scoring:
  - Component 1: All 5 restaurants have non-empty Address values (0.35 pts)
  - Component 2: All 5 restaurants have non-empty Website values (0.30 pts)
  - Component 3: All 5 restaurants have non-empty Phone values (0.35 pts)

The initial state has restaurant names pre-filled but Address/Website/Phone all empty.
This script verifies that the agent has successfully looked up and entered the missing data.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_002'

EXPECTED_RESTAURANTS = [
    'Harutaka',
    'Sukiyabashi Jiro Honten',
    'Sushi Saito',
    'Sushi Yoshitake',
    'Sushi Sho',
]

FILE_PATH = f'{WORKDIR}/Desktop/SUSHI_SPOTS.xlsx'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires filling in Address, Website, and Phone for 5 Tokyo sushi
    restaurants that were initially empty in the spreadsheet.
    """
    total_score = 0.0

    # Load the workbook — fail fast if we can't even open it
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active sheet (expected to be 'Sushi Spots')
    try:
        ws = wb.active
        print(f"INFO: Active sheet: '{ws.title}', rows: {ws.max_row}, cols: {ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition gate: verify basic structure ---
    # Check that headers exist (Name, Address, Website, Phone in row 1)
    try:
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        if not all(h is not None for h in headers):
            print(f"CRITICAL: Header row missing or incomplete: {headers}")
            print("REWARD: 0.0")
            return 0.0

        # Check that restaurant names are still in column A rows 2-6
        names_in_file = [ws.cell(row=r, column=1).value for r in range(2, 7)]
        found_restaurants = [n for n in names_in_file if n is not None]
        if len(found_restaurants) < 5:
            print(f"CRITICAL: Expected 5 restaurant names in column A, found: {found_restaurants}")
            print("REWARD: 0.0")
            return 0.0
        print(f"INFO: Restaurants found: {found_restaurants}")
    except Exception as e:
        print(f"CRITICAL: Structure check failed: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Build a mapping from row index to restaurant name for reporting
    restaurant_rows = {}
    for r in range(2, 7):
        name = ws.cell(row=r, column=1).value
        if name is not None:
            restaurant_rows[r] = str(name).strip()

    # --- Component 1: All 5 restaurants have non-empty Address values (0.35 pts) ---
    # Addresses are in column B (column index 2)
    # In initial_env: all address cells are None
    # In golden_env: all address cells have values like "6F, 8-3-1 Ginza, Chuo City, Tokyo 104-0061"
    try:
        address_fill_count = 0
        address_details = []
        for r, name in restaurant_rows.items():
            addr = ws.cell(row=r, column=2).value
            if addr is not None and str(addr).strip() != '':
                address_fill_count += 1
                address_details.append(f"  Row {r} ({name}): '{str(addr)[:60]}'")
            else:
                address_details.append(f"  Row {r} ({name}): EMPTY")

        if address_fill_count == 5:
            print(f"PASS: Component 1 — All 5 addresses filled (0.35 pts)")
            for d in address_details:
                print(d)
            total_score += 0.35
        elif address_fill_count > 0:
            partial = round(address_fill_count / 5 * 0.35, 4)
            print(f"PARTIAL: Component 1 — {address_fill_count}/5 addresses filled ({partial} pts)")
            for d in address_details:
                print(d)
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No addresses filled (0.0 pts)")
            for d in address_details:
                print(d)
    except Exception as e:
        print(f"ERROR: Component 1 (Address check) — {e}")

    # --- Component 2: All 5 restaurants have non-empty Website values (0.30 pts) ---
    # Websites are in column C (column index 3)
    # In initial_env: all website cells are None
    # In golden_env: all website cells have URLs like "https://harutaka.jp"
    try:
        website_fill_count = 0
        website_details = []
        for r, name in restaurant_rows.items():
            website = ws.cell(row=r, column=3).value
            if website is not None and str(website).strip() != '':
                website_fill_count += 1
                website_details.append(f"  Row {r} ({name}): '{str(website)[:80]}'")
            else:
                website_details.append(f"  Row {r} ({name}): EMPTY")

        if website_fill_count == 5:
            print(f"PASS: Component 2 — All 5 websites filled (0.30 pts)")
            for d in website_details:
                print(d)
            total_score += 0.30
        elif website_fill_count > 0:
            partial = round(website_fill_count / 5 * 0.30, 4)
            print(f"PARTIAL: Component 2 — {website_fill_count}/5 websites filled ({partial} pts)")
            for d in website_details:
                print(d)
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No websites filled (0.0 pts)")
            for d in website_details:
                print(d)
    except Exception as e:
        print(f"ERROR: Component 2 (Website check) — {e}")

    # --- Component 3: All 5 restaurants have non-empty Phone values (0.35 pts) ---
    # Phone numbers are in column D (column index 4)
    # In initial_env: all phone cells are None
    # In golden_env: all phone cells have values like "03-3573-1144"
    try:
        phone_fill_count = 0
        phone_details = []
        for r, name in restaurant_rows.items():
            phone = ws.cell(row=r, column=4).value
            if phone is not None and str(phone).strip() != '':
                phone_fill_count += 1
                phone_details.append(f"  Row {r} ({name}): '{str(phone)[:40]}'")
            else:
                phone_details.append(f"  Row {r} ({name}): EMPTY")

        if phone_fill_count == 5:
            print(f"PASS: Component 3 — All 5 phone numbers filled (0.35 pts)")
            for d in phone_details:
                print(d)
            total_score += 0.35
        elif phone_fill_count > 0:
            partial = round(phone_fill_count / 5 * 0.35, 4)
            print(f"PARTIAL: Component 3 — {phone_fill_count}/5 phones filled ({partial} pts)")
            for d in phone_details:
                print(d)
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No phone numbers filled (0.0 pts)")
            for d in phone_details:
                print(d)
    except Exception as e:
        print(f"ERROR: Component 3 (Phone check) — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: run against the canonical artifact path
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
