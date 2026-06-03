"""
Reward Script: Build a reusable monthly expense report template for employee reimbursements.
Task ID: calc_gen_template_037
Domain: libreoffice_calc

Scoring rubric (total = 1.0):
  Component 1: Header row A1 has company title, merged A1:G1, bold (0.15)
  Component 2: Column headers in Row 5 exist (Date/Description/Category/Miles/Mileage$/Receipt$/Total) (0.15)
  Component 3: Mileage formulas =D*0.67 in E6:E20 (IRS rate) (0.20)
  Component 4: Total formulas =E*+F* in G6:G20 (0.10)
  Component 5: TOTALS row at row 21 with SUM formulas for columns D/E/F/G (0.15)
  Component 6: Category dropdown in C6:C20 with valid categories (0.10)
  Component 7: Manager approval section in rows 23-25 (0.05)
  Component 8: Print settings: print area A1:G25, fit to 1 page wide, header rows repeated (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_template_037'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet 'ExpenseReport' exists
    if 'ExpenseReport' not in wb.sheetnames:
        print("FAIL: Sheet 'ExpenseReport' not found")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['ExpenseReport']

    # Component 1: Header row — A1 contains company title, merged A1:G1, bold (0.15 points)
    try:
        a1_value = ws['A1'].value
        # Check value is a non-empty string containing expected content
        has_title = (
            a1_value is not None and
            isinstance(a1_value, str) and
            len(a1_value.strip()) > 0 and
            'expense report' in a1_value.lower()
        )

        # Check merged — B1 through G1 should be MergedCell
        is_merged = isinstance(ws['B1'], MergedCell) or 'A1:G1' in [str(m) for m in ws.merged_cells.ranges]

        # Check bold
        is_bold = ws['A1'].font.bold is True

        if has_title and is_merged and is_bold:
            print(f"PASS: Component 1 — Header row A1 has title '{a1_value}', merged A1:G1, bold (0.15 pts)")
            total_score += 0.15
        else:
            details = []
            if not has_title:
                details.append(f"A1 value '{a1_value}' does not contain 'expense report'")
            if not is_merged:
                details.append("A1:G1 not merged")
            if not is_bold:
                details.append("A1 not bold")
            print(f"FAIL: Component 1 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column headers in Row 5 (0.15 points)
    try:
        expected_headers = ['Date', 'Description', 'Category', 'Miles', 'Mileage $', 'Receipt $', 'Total']
        actual_headers = [ws.cell(row=5, column=col).value for col in range(1, 8)]

        headers_match = all(
            actual_headers[i] is not None and
            str(actual_headers[i]).strip().lower() == expected_headers[i].lower()
            for i in range(len(expected_headers))
        )

        if headers_match:
            print(f"PASS: Component 2 — Row 5 has all 7 column headers: {actual_headers} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Row 5 headers expected {expected_headers}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Mileage formulas =D*0.67 in E6:E20 (IRS rate) (0.20 points)
    try:
        mileage_correct = 0
        total_rows = 15  # rows 6-20

        for row in range(6, 21):
            cell_val = ws.cell(row=row, column=5).value  # column E
            if cell_val is not None and isinstance(cell_val, str):
                # Check for IRS rate formula like =D6*0.67
                formula = cell_val.strip().upper().replace(' ', '')
                if f'=D{row}*0.67' in formula or formula == f'=D{row}*0.67':
                    mileage_correct += 1

        if mileage_correct == total_rows:
            print(f"PASS: Component 3 — All {total_rows} mileage formulas =D*0.67 present in E6:E20 (0.20 pts)")
            total_score += 0.20
        elif mileage_correct > 0:
            partial = round(0.20 * mileage_correct / total_rows, 3)
            print(f"PARTIAL: Component 3 — {mileage_correct}/{total_rows} mileage formulas correct (partial {partial} pts awarded as 0.0 — all must be present)")
            print(f"FAIL: Component 3 — only {mileage_correct}/{total_rows} mileage formula rows correct")
        else:
            print(f"FAIL: Component 3 — No mileage formulas =D*0.67 found in E6:E20")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Total formulas =E*+F* in G6:G20 (0.10 points)
    try:
        total_correct = 0
        total_rows = 15  # rows 6-20

        for row in range(6, 21):
            cell_val = ws.cell(row=row, column=7).value  # column G
            if cell_val is not None and isinstance(cell_val, str):
                formula = cell_val.strip().upper().replace(' ', '')
                if f'=E{row}+F{row}' in formula or formula == f'=E{row}+F{row}':
                    total_correct += 1

        if total_correct == total_rows:
            print(f"PASS: Component 4 — All {total_rows} total formulas =E*+F* present in G6:G20 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — only {total_correct}/{total_rows} total formula rows correct in G6:G20")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: TOTALS row at row 21 with SUM formulas for D, E, F, G columns (0.15 points)
    try:
        totals_label = ws.cell(row=21, column=1).value
        has_label = totals_label is not None and 'total' in str(totals_label).lower()

        sum_d = ws.cell(row=21, column=4).value  # D21
        sum_e = ws.cell(row=21, column=5).value  # E21
        sum_f = ws.cell(row=21, column=6).value  # F21
        sum_g = ws.cell(row=21, column=7).value  # G21

        def is_sum_formula(val, col_letter):
            if not val or not isinstance(val, str):
                return False
            v = val.strip().upper().replace(' ', '')
            return f'SUM({col_letter}6:{col_letter}20)' in v or f'SUM({col_letter}6:{col_letter}20)' in v

        has_sum_d = is_sum_formula(sum_d, 'D')
        has_sum_e = is_sum_formula(sum_e, 'E')
        has_sum_f = is_sum_formula(sum_f, 'F')
        has_sum_g = is_sum_formula(sum_g, 'G')

        if has_label and has_sum_d and has_sum_e and has_sum_f and has_sum_g:
            print(f"PASS: Component 5 — TOTALS row at row 21 with SUM formulas for D/E/F/G (0.15 pts)")
            total_score += 0.15
        else:
            details = []
            if not has_label:
                details.append(f"A21 label is '{totals_label}' (expected 'TOTALS')")
            if not has_sum_d:
                details.append(f"D21='{sum_d}' (expected =SUM(D6:D20))")
            if not has_sum_e:
                details.append(f"E21='{sum_e}' (expected =SUM(E6:E20))")
            if not has_sum_f:
                details.append(f"F21='{sum_f}' (expected =SUM(F6:F20))")
            if not has_sum_g:
                details.append(f"G21='{sum_g}' (expected =SUM(G6:G20))")
            print(f"FAIL: Component 5 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Category dropdown in C6:C20 with Meals, Travel, Hotel, Mileage, Supplies, Other (0.10 points)
    try:
        expected_categories = ['Meals', 'Travel', 'Hotel', 'Mileage', 'Supplies', 'Other']

        # Scan data validations for a list dropdown in column C covering the data entry rows
        dropdown_dvs = [
            dv for dv in ws.data_validations.dataValidation
            if dv.type == 'list' and ('C6' in str(dv.sqref).upper() or 'C' in str(dv.sqref).upper())
        ]
        found_dropdown = len(dropdown_dvs) > 0

        # Check whether any of the found dropdowns has at least 5/6 required categories
        matching_dv_count = 0
        if found_dropdown and dropdown_dvs[0].formula1:
            formula_clean = dropdown_dvs[0].formula1.replace('"', '').strip()
            cats_in_formula = [c.strip() for c in formula_clean.split(',')]
            matching_dv_count = sum(
                1 for c in expected_categories
                if any(c.lower() == x.lower() for x in cats_in_formula)
            )

        if found_dropdown and matching_dv_count >= 5:
            print(f"PASS: Component 6 — Category dropdown in C6:C20 with correct categories (0.10 pts)")
            total_score += 0.10
        elif found_dropdown:
            print(f"FAIL: Component 6 — Dropdown found but only {matching_dv_count}/{len(expected_categories)} categories matched")
        else:
            print(f"FAIL: Component 6 — No list dropdown found in column C for data entry rows")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Manager approval section in rows 23-25 (0.05 points)
    try:
        a23 = ws.cell(row=23, column=1).value
        a24 = ws.cell(row=24, column=1).value
        a25 = ws.cell(row=25, column=1).value

        has_approval_header = a23 is not None and 'approval' in str(a23).lower()
        has_approved_by = a24 is not None and 'approved' in str(a24).lower()
        has_comments = a25 is not None and 'comment' in str(a25).lower()

        if has_approval_header and has_approved_by and has_comments:
            print(f"PASS: Component 7 — Manager approval section present in rows 23-25 (0.05 pts)")
            total_score += 0.05
        else:
            details = []
            if not has_approval_header:
                details.append(f"A23='{a23}' (expected manager approval header)")
            if not has_approved_by:
                details.append(f"A24='{a24}' (expected 'Approved by:')")
            if not has_comments:
                details.append(f"A25='{a25}' (expected 'Comments:')")
            print(f"FAIL: Component 7 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: Print settings — print area A1:G25, fit to 1 page wide, header row repeated (0.10 points)
    try:
        print_area = ws.print_area
        # print_area may include sheet name prefix like "'ExpenseReport'!$A$1:$G$25"
        # The format uses dollar signs, so check for $A$1 and $G$25 (case-insensitive)
        pa_str = str(print_area).upper() if print_area else ''
        has_print_area = (print_area is not None and
                          ('$A$1' in pa_str or 'A1' in pa_str.replace('$', '')) and
                          ('$G$25' in pa_str or 'G25' in pa_str.replace('$', '')))

        try:
            fit_width = ws.page_setup.fitToWidth
            has_fit_to_page = fit_width == 1
        except Exception:
            has_fit_to_page = False

        try:
            rows_to_repeat = ws.print_title_rows
            has_rows_repeat = rows_to_repeat is not None and '$1' in str(rows_to_repeat)
        except Exception:
            has_rows_repeat = False

        if has_print_area and has_fit_to_page and has_rows_repeat:
            print(f"PASS: Component 8 — Print area A1:G25, fitToWidth=1, header rows repeated (0.10 pts)")
            total_score += 0.10
        else:
            details = []
            if not has_print_area:
                details.append(f"print_area='{print_area}' (expected $A$1:$G$25)")
            if not has_fit_to_page:
                details.append(f"fitToWidth={ws.page_setup.fitToWidth} (expected 1)")
            if not has_rows_repeat:
                details.append(f"print_title_rows='{ws.print_title_rows}' (expected $1:$5)")
            print(f"FAIL: Component 8 — {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
