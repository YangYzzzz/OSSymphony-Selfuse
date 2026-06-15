"""
Initial Setup: Wide report with 16 columns (A-P) and 200 rows, no print scaling or repeating headers.
Task ID: calc_tbl_047
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- Headers (A-P, 16 columns) ---
    headers = [
        "Employee ID", "Full Name", "Department", "Job Title",
        "Office Location", "Hire Date", "Base Salary", "Bonus",
        "Total Comp", "Manager", "Email", "Phone Extension",
        "Performance Rating", "Projects Completed", "Training Hours", "Notes"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths to make the report realistically wide ---
    col_widths = {
        "A": 12, "B": 22, "C": 18, "D": 24, "E": 18, "F": 14,
        "G": 14, "H": 12, "I": 14, "J": 22, "K": 28, "L": 14,
        "M": 16, "N": 18, "O": 16, "P": 30,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1 height
    ws.row_dimensions[1].height = 32

    # --- Data generation (rows 2-200) ---
    first_names = [
        "Sarah", "Marcus", "Elena", "David", "Priya", "James", "Lin",
        "Fatima", "Carlos", "Yuki", "Olga", "Amir", "Zara", "Thomas",
        "Keiko", "Roberto", "Amara", "Wei", "Ingrid", "Samuel"
    ]
    last_names = [
        "Chen", "Johnson", "Rodriguez", "Kim", "Patel", "O'Brien",
        "Nakamura", "Hassan", "Morales", "Dubois", "Kowalski", "Singh",
        "Okafor", "Larsson", "Tanaka", "Ferreira", "Diallo", "Zhang",
        "Johansson", "Osei"
    ]
    departments = [
        "Engineering", "Marketing", "Finance", "Human Resources",
        "Operations", "Sales", "Legal", "Product", "Design", "IT Support"
    ]
    titles = [
        "Software Engineer", "Senior Analyst", "Project Manager",
        "Team Lead", "Director", "Coordinator", "Specialist",
        "VP Operations", "Associate", "Principal Engineer",
        "Data Scientist", "UX Designer", "Account Manager"
    ]
    offices = [
        "New York", "San Francisco", "London", "Tokyo", "Singapore",
        "Berlin", "Sydney", "Toronto", "Mumbai", "Sao Paulo"
    ]
    managers = [
        "Jennifer Walsh", "Michael Torres", "Anna Bergstrom",
        "Raj Krishnan", "Sophie Martin", "Daniel Wright"
    ]
    notes_options = [
        "On track for promotion", "Completed leadership training",
        "Transferred from London office", "Working on Project Atlas",
        "Mentoring 2 junior engineers", "Led Q3 product launch",
        "Remote worker - approved", "Returning from sabbatical",
        "Cross-functional team lead", "Patent pending - automation system",
        "Exceeded quarterly targets", "Bilingual - English/Mandarin",
        "Six Sigma certified", "MBA in progress", ""
    ]

    for r in range(2, 201):
        emp_id = 10000 + r - 1
        fname = random.choice(first_names)
        lname = random.choice(last_names)
        full_name = f"{fname} {lname}"
        dept = random.choice(departments)
        title = random.choice(titles)
        office = random.choice(offices)
        year = random.randint(2015, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        hire_date = f"{year}-{month:02d}-{day:02d}"
        base_salary = random.randint(55000, 185000)
        bonus = round(base_salary * random.uniform(0.05, 0.25), 2)
        total_comp = base_salary + bonus
        mgr = random.choice(managers)
        email = f"{fname.lower()}.{lname.lower()}@globalcorp.com"
        ext = random.randint(1000, 9999)
        perf_rating = round(random.uniform(2.5, 5.0), 1)
        projects = random.randint(1, 18)
        training_hrs = random.randint(8, 120)
        note = random.choice(notes_options)

        row_data = [
            emp_id, full_name, dept, title, office, hire_date,
            base_salary, bonus, total_comp, mgr, email, ext,
            perf_rating, projects, training_hrs, note
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (7, 8, 9):  # salary columns
                cell.number_format = '$#,##0.00'
            elif c == 6:  # date
                cell.number_format = 'yyyy-mm-dd'

    # No print scaling, no repeating headers -- that is the task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
