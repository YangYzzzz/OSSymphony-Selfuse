"""
Initial Setup: Budget spreadsheet with named ranges (Expenses C2:C10, Revenue B2:B10, NetProfit D2:D10)
Task ID: calc_cop_named_range_003
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_named_range_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Budget ---
    ws = wb.active
    ws.title = 'Budget'

    # Headers
    headers = ['Month', 'Revenue', 'Expenses', 'NetProfit']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows 2-10 (original data for named ranges C2:C10, B2:B10, D2:D10)
    data = [
        ['January',  42500, 31200, 11300],
        ['February', 38700, 28900,  9800],
        ['March',    51200, 35600, 15600],
        ['April',    47800, 33100, 14700],
        ['May',      55300, 38400, 16900],
        ['June',     49600, 34700, 14900],
        ['July',     53100, 37200, 15900],
        ['August',   48400, 33800, 14600],
        ['September',46900, 32500, 14400],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Rows 11-15: new expense data added (but named range Expenses still only covers C2:C10)
    new_data = [
        ['October',  44100, 3400,  None],
        ['November', 39800, 2100,  None],
        ['December', 58200, 5600,  None],
        ['Q4-Adj1',  None,  1800,  None],
        ['Q4-Adj2',  None,  4200,  None],
    ]
    for r, row_data in enumerate(new_data, 11):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)

    # --- Define Named Ranges ---
    # Expenses: C2:C10 (to be expanded to C2:C15 by the task)
    dn_expenses = DefinedName('Expenses', attr_text="Budget!$C$2:$C$10")
    wb.defined_names.add(dn_expenses)

    # Revenue: B2:B10
    dn_revenue = DefinedName('Revenue', attr_text="Budget!$B$2:$B$10")
    wb.defined_names.add(dn_revenue)

    # NetProfit: D2:D10
    dn_netprofit = DefinedName('NetProfit', attr_text="Budget!$D$2:$D$10")
    wb.defined_names.add(dn_netprofit)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Named ranges created:')
    print('  - Expenses: Budget!$C$2:$C$10')
    print('  - Revenue: Budget!$B$2:$B$10')
    print('  - NetProfit: Budget!$D$2:$D$10')
    print('Rows 11-15 in column C contain: 3400, 2100, 5600, 1800, 4200')


create_initial()
