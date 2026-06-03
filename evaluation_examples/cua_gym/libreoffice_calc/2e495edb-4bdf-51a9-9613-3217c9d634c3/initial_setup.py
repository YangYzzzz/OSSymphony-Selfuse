"""
Initial Setup: Project Budget Tracker with gantt-like status section
Task ID: calc_gsd_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Tracker"

    # Row 1: intentionally left blank (agent will merge and format)

    # Row 2: Headers (plain, no formatting)
    headers = [
        "Task", "Owner", "Start Date", "End Date",
        "Budget", "Actual Cost", "Variance", "% Complete", "Status"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # Rows 3-27: 25 project tasks with realistic data
    tasks = [
        ["Requirements Gathering", "Sarah Chen", "2025-01-06", "2025-01-24", 12000, 11500, 500, 100, "Complete"],
        ["Stakeholder Interviews", "Sarah Chen", "2025-01-13", "2025-01-31", 8000, 8200, -200, 100, "Complete"],
        ["System Architecture Design", "Marcus Johnson", "2025-02-03", "2025-02-28", 25000, 23800, 1200, 100, "Complete"],
        ["Database Schema Design", "Marcus Johnson", "2025-02-10", "2025-02-21", 15000, 14500, 500, 100, "Complete"],
        ["UI/UX Wireframes", "Priya Patel", "2025-02-17", "2025-03-07", 18000, 19200, -1200, 95, "In Progress"],
        ["Frontend Development - Phase 1", "Wei Zhang", "2025-03-03", "2025-04-04", 45000, 42000, 3000, 90, "In Progress"],
        ["Backend API Development", "David Kim", "2025-03-03", "2025-04-11", 52000, 51500, 500, 85, "In Progress"],
        ["Authentication Module", "David Kim", "2025-03-17", "2025-03-28", 14000, 13800, 200, 80, "In Progress"],
        ["Payment Integration", "Aisha Rahman", "2025-03-24", "2025-04-18", 22000, 24500, -2500, 70, "In Progress"],
        ["Third-party API Connectors", "Wei Zhang", "2025-04-07", "2025-04-25", 16000, 15000, 1000, 65, "In Progress"],
        ["Data Migration Scripts", "Elena Volkov", "2025-04-14", "2025-05-02", 20000, 21000, -1000, 60, "In Progress"],
        ["Unit Testing - Backend", "David Kim", "2025-04-21", "2025-05-09", 18000, 17500, 500, 55, "In Progress"],
        ["Unit Testing - Frontend", "Wei Zhang", "2025-04-28", "2025-05-16", 16000, 16800, -800, 50, "In Progress"],
        ["Integration Testing", "Aisha Rahman", "2025-05-05", "2025-05-23", 24000, 25000, -1000, 45, "At Risk"],
        ["Performance Testing", "Elena Volkov", "2025-05-12", "2025-05-30", 14000, 13500, 500, 40, "At Risk"],
        ["Security Audit", "Marcus Johnson", "2025-05-19", "2025-06-06", 30000, 32000, -2000, 35, "At Risk"],
        ["User Acceptance Testing", "Priya Patel", "2025-05-26", "2025-06-13", 12000, 11000, 1000, 30, "Planned"],
        ["Documentation - Technical", "Elena Volkov", "2025-06-02", "2025-06-20", 10000, 9500, 500, 25, "Planned"],
        ["Documentation - User Guide", "Priya Patel", "2025-06-09", "2025-06-27", 8000, 7800, 200, 20, "Planned"],
        ["Training Materials", "Sarah Chen", "2025-06-16", "2025-07-04", 11000, 12000, -1000, 15, "Planned"],
        ["Deployment - Staging", "David Kim", "2025-06-23", "2025-07-04", 9000, 8500, 500, 10, "Planned"],
        ["Deployment - Production", "Marcus Johnson", "2025-07-07", "2025-07-18", 15000, 14000, 1000, 5, "Not Started"],
        ["Post-Launch Monitoring", "Aisha Rahman", "2025-07-14", "2025-08-01", 18000, 17500, 500, 0, "Not Started"],
        ["Bug Fix Sprint", "Wei Zhang", "2025-07-21", "2025-08-08", 20000, 19000, 1000, 0, "Not Started"],
        ["Project Retrospective", "Sarah Chen", "2025-08-11", "2025-08-15", 5000, 4800, 200, 0, "Not Started"],
    ]

    for r, row_data in enumerate(tasks, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
