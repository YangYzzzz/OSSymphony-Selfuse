"""
Reward Script: Project budget tracker with formatted header, conditional formatting,
               currency formatting, and summary section with variance analysis.
Task ID: calc_gsd_030
Domain: libreoffice_calc
Scoring:
  Component 1: Merged header row A1:I1 with title, formatting (0.20)
  Component 2: Row 2 header styling - bold, blue bg, white text (0.15)
  Component 3: Currency formatting on E3:F27 (Budget, Actual Cost) (0.15)
  Component 4: Currency formatting on G3:G27 (Variance) with red negatives (0.15)
  Component 5: Icon set conditional formatting on H3:H27 (0.15)
  Component 6: Summary rows 29-31 with SUM formulas and bold labels (0.20)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_030'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    from openpyxl.cell.cell import MergedCell

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Budget Tracker' not in wb.sheetnames:
        print("CRITICAL: 'Budget Tracker' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget Tracker']

    # =====================================================================
    # Component 1: Merged header A1:I1 with title, formatting (0.20 points)
    # =====================================================================
    try:
        # Check merge exists
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'I1' in r for r in merged_ranges)

        a1 = ws['A1']
        has_title = (a1.value is not None and
                     'project apollo' in str(a1.value).lower() and
                     'budget' in str(a1.value).lower())
        has_bold = (a1.font.bold == True)
        has_large_font = (a1.font.size is not None and a1.font.size >= 14)

        # Check dark blue background (FF1F3864)
        try:
            fill_rgb = a1.fill.fgColor.rgb
            has_dark_bg = (fill_rgb is not None and
                           fill_rgb.upper() in ('FF1F3864', '001F3864', '1F3864'))
        except Exception:
            has_dark_bg = False

        # Check white font color
        try:
            font_rgb = a1.font.color.rgb
            has_white_font = (font_rgb is not None and
                              font_rgb.upper().endswith('FFFFFF'))
        except Exception:
            has_white_font = False

        sub_checks = [has_merge, has_title, has_bold, has_large_font, has_dark_bg, has_white_font]
        passed = sum(sub_checks)

        if passed >= 4:
            # At minimum: merge + title + bold + font size
            score_1 = 0.20
        elif passed >= 2:
            score_1 = 0.10
        else:
            score_1 = 0.0

        if score_1 > 0:
            print(f"PASS: Component 1 — Merged header ({passed}/6 sub-checks) ({score_1} pts)")
            total_score += score_1
        else:
            print(f"FAIL: Component 1 — Merged header: merge={has_merge}, title={has_title}, bold={has_bold}, font_size={has_large_font}, dark_bg={has_dark_bg}, white_font={has_white_font}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =====================================================================
    # Component 2: Row 2 header styling (0.15 points)
    # =====================================================================
    try:
        bold_count = 0
        blue_bg_count = 0
        white_font_count = 0

        for col in range(1, 10):
            c = ws.cell(row=2, column=col)
            if c.font.bold:
                bold_count += 1
            try:
                fill_rgb = c.fill.fgColor.rgb
                if fill_rgb and fill_rgb.upper() in ('FF4472C4', '004472C4', '4472C4'):
                    blue_bg_count += 1
            except:
                pass
            try:
                font_rgb = c.font.color.rgb
                if font_rgb and font_rgb.upper().endswith('FFFFFF'):
                    white_font_count += 1
            except:
                pass

        # Need majority of 9 columns to have bold + blue bg
        if bold_count >= 7 and blue_bg_count >= 7:
            print(f"PASS: Component 2 — Row 2 styled: bold={bold_count}/9, blue_bg={blue_bg_count}/9, white_font={white_font_count}/9 (0.15 pts)")
            total_score += 0.15
        elif bold_count >= 5:
            print(f"PARTIAL: Component 2 — Row 2 partially styled: bold={bold_count}/9, blue_bg={blue_bg_count}/9 (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2 — Row 2 not styled: bold={bold_count}/9, blue_bg={blue_bg_count}/9")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =====================================================================
    # Component 3: Currency formatting on E3:F27 (0.15 points)
    # =====================================================================
    try:
        currency_count = 0
        total_cells = 0
        for row in range(3, 28):
            for col in [5, 6]:  # E, F
                c = ws.cell(row=row, column=col)
                if c.value is not None:
                    total_cells += 1
                    fmt = c.number_format or ''
                    if '$' in fmt or 'USD' in fmt.upper():
                        currency_count += 1

        if total_cells > 0:
            ratio = currency_count / total_cells
        else:
            ratio = 0

        if ratio >= 0.8:
            print(f"PASS: Component 3 — Currency on E:F columns: {currency_count}/{total_cells} ({ratio:.0%}) (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.4:
            print(f"PARTIAL: Component 3 — Currency on E:F columns: {currency_count}/{total_cells} ({ratio:.0%}) (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 — Currency on E:F columns: {currency_count}/{total_cells}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =====================================================================
    # Component 4: Variance column G3:G27 currency with red negatives (0.15 points)
    # =====================================================================
    try:
        variance_currency = 0
        variance_red_neg = 0
        variance_total = 0

        for row in range(3, 28):
            c = ws.cell(row=row, column=7)  # G column
            if c.value is not None:
                variance_total += 1
                fmt = c.number_format or ''
                if '$' in fmt:
                    variance_currency += 1
                # Check for red/parentheses negative format
                if 'Red' in fmt or '(' in fmt:
                    variance_red_neg += 1

        if variance_total > 0:
            curr_ratio = variance_currency / variance_total
            red_ratio = variance_red_neg / variance_total
        else:
            curr_ratio = 0
            red_ratio = 0

        if curr_ratio >= 0.8 and red_ratio >= 0.8:
            print(f"PASS: Component 4 — Variance formatting: currency={variance_currency}/{variance_total}, red_neg={variance_red_neg}/{variance_total} (0.15 pts)")
            total_score += 0.15
        elif curr_ratio >= 0.8:
            print(f"PARTIAL: Component 4 — Variance has currency but no red negatives (0.10 pts)")
            total_score += 0.10
        elif curr_ratio >= 0.4:
            print(f"PARTIAL: Component 4 — Some variance formatting (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — Variance not formatted: currency={variance_currency}/{variance_total}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =====================================================================
    # Component 5: Icon set conditional formatting on H3:H27 (0.15 points)
    # =====================================================================
    try:
        icon_set_found = 0  # count of iconSet rules found
        icon_range_match = 0  # count matching H column

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'iconSet':
                    icon_set_found += 1
                    # Check if range covers H3:H27 (or close)
                    if 'H3' in cf_range or 'H' in cf_range:
                        icon_range_match += 1

        if icon_set_found > 0 and icon_range_match > 0:
            print(f"PASS: Component 5 — Icon set CF on H column (0.15 pts)")
            total_score += 0.15
        elif icon_set_found > 0:
            print(f"PARTIAL: Component 5 — Icon set CF found but range unclear (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No icon set conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =====================================================================
    # Component 6: Summary rows 29-31 with SUM formulas and bold labels (0.20 points)
    # =====================================================================
    try:
        summary_score = 0.0

        # Check for summary labels in column A (rows 29-31)
        labels_found = 0
        bold_labels = 0
        for row in range(29, 32):
            label = ws.cell(row=row, column=1).value
            is_bold = ws.cell(row=row, column=1).font.bold
            if label and any(kw in str(label).lower() for kw in ['total', 'budget', 'actual', 'variance']):
                labels_found += 1
                if is_bold:
                    bold_labels += 1

        # Check for SUM formulas in the summary rows
        sum_formulas_found = 0
        expected_sums = {
            (29, 5): 'E',   # Total Budget -> SUM(E3:E27)
            (30, 6): 'F',   # Total Actual -> SUM(F3:F27)
            (31, 7): 'G',   # Total Variance -> SUM(G3:G27)
        }

        for (row, col), col_letter in expected_sums.items():
            c = ws.cell(row=row, column=col)
            val = c.value
            if val and isinstance(val, str) and '=SUM' in val.upper():
                # Verify it references the correct column range
                if col_letter in val.upper():
                    sum_formulas_found += 1
                    print(f"  Found SUM formula at row {row}, col {col}: {val}")
                else:
                    print(f"  SUM formula at row {row}, col {col} references wrong column: {val}")
            else:
                print(f"  No SUM formula at row {row}, col {col}: {repr(val)}")

        if labels_found >= 3 and bold_labels >= 3 and sum_formulas_found >= 3:
            summary_score = 0.20
        elif labels_found >= 2 and sum_formulas_found >= 2:
            summary_score = 0.15
        elif sum_formulas_found >= 1:
            summary_score = 0.07
        else:
            summary_score = 0.0

        if summary_score > 0:
            print(f"PASS: Component 6 — Summary section: labels={labels_found}/3, bold={bold_labels}/3, sums={sum_formulas_found}/3 ({summary_score} pts)")
            total_score += summary_score
        else:
            print(f"FAIL: Component 6 — Summary section: labels={labels_found}/3, bold={bold_labels}/3, sums={sum_formulas_found}/3")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
