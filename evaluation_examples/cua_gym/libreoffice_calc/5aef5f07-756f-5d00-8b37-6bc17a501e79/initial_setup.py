"""
Initial Setup: Restaurant menu and ingredient cost calculator
Task ID: calc_grs_047
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ── Styles ──
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # ===================================================================
    # Sheet1: Ingredient Inventory
    # ===================================================================
    ws1 = wb.active
    ws1.title = "Ingredient Inventory"

    headers1 = ["Ingredient Name", "Unit", "Cost per Unit", "Stock on Hand"]
    for c, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=c, value=h)
    style_header(ws1, 1, len(headers1))

    ingredients = [
        ["Chicken Breast",   "kg",   12.50, 45],
        ["Salmon Fillet",    "kg",   28.00, 18],
        ["Ground Beef",      "kg",   15.75, 30],
        ["Pasta (Penne)",    "pack",  2.40, 60],
        ["Olive Oil",        "L",    14.00, 12],
        ["Heavy Cream",      "L",     6.80, 20],
        ["Parmesan Cheese",  "kg",   32.00,  8],
        ["Fresh Tomatoes",   "kg",    4.50, 25],
        ["Mixed Greens",     "kg",    8.90, 15],
        ["Garlic",           "kg",   11.00,  5],
        ["Shrimp (Peeled)",  "kg",   26.50, 12],
        ["Mushrooms",        "kg",    9.20, 10],
        ["Balsamic Vinegar", "L",    18.00,  4],
        ["Bread Rolls",      "pack",  3.20, 40],
        ["Butter",           "kg",   10.50, 15],
    ]

    for r, row_data in enumerate(ingredients, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:
                cell.number_format = '$#,##0.00'
            if c == 4:
                cell.number_format = '0'

    # Data validation for Unit column
    dv = DataValidation(
        type="list",
        formula1='"kg,L,unit,pack"',
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "Please select a valid unit"
    dv.errorTitle = "Invalid Unit"
    dv.prompt = "Select unit of measurement"
    dv.promptTitle = "Unit"
    dv.add(f"B2:B{len(ingredients) + 1}")
    ws1.add_data_validation(dv)

    ws1.column_dimensions["A"].width = 22
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 16
    ws1.freeze_panes = "A2"

    # ===================================================================
    # Sheet2: Recipe Book
    # ===================================================================
    ws2 = wb.create_sheet("Recipe Book")

    # Columns: Menu Item, Selling Price, then one column per ingredient (quantity in units)
    ingredient_names = [row[0] for row in ingredients]
    headers2 = ["Menu Item", "Selling Price"] + ingredient_names
    for c, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=c, value=h)
    style_header(ws2, 1, len(headers2))

    # Menu items: [Name, Selling Price, ingredient quantities...]
    # Quantities represent how much of each ingredient is used per serving
    # Order matches ingredient list above
    menu_items = [
        ["Grilled Chicken Salad",    16.50, 0.20, 0,    0,    0,    0.02, 0,    0.03, 0.08, 0.10, 0,    0,    0,    0.01, 0,    0.01],
        ["Pan-Seared Salmon",        24.00, 0,    0.18, 0,    0,    0.02, 0.05, 0,    0.06, 0,    0.01, 0,    0.04, 0,    0,    0.02],
        ["Beef Bolognese Pasta",     18.00, 0,    0,    0.22, 0.15, 0.02, 0,    0.03, 0.10, 0,    0.02, 0,    0,    0,    0,    0.01],
        ["Shrimp Alfredo",           22.50, 0,    0,    0,    0.15, 0.01, 0.10, 0.04, 0,    0,    0.01, 0.15, 0,    0,    0,    0.02],
        ["Mushroom Risotto",         17.00, 0,    0,    0,    0,    0.03, 0.08, 0.05, 0,    0,    0.01, 0,    0.12, 0,    0,    0.03],
        ["Classic Caesar Salad",     13.50, 0.15, 0,    0,    0,    0.02, 0,    0.05, 0,    0.12, 0.01, 0,    0,    0,    0.05, 0],
        ["Garlic Butter Shrimp",     21.00, 0,    0,    0,    0,    0.02, 0,    0,    0.05, 0,    0.03, 0.20, 0,    0,    0.02, 0.03],
        ["Tomato Bruschetta",        11.00, 0,    0,    0,    0,    0.02, 0,    0,    0.15, 0,    0.02, 0,    0,    0.02, 0.10, 0.01],
    ]

    for r, row_data in enumerate(menu_items, 2):
        name = row_data[0]
        price = row_data[1]
        quantities = row_data[2:]
        ws2.cell(row=r, column=1, value=name).border = thin_border
        cell_price = ws2.cell(row=r, column=2, value=price)
        cell_price.number_format = '$#,##0.00'
        cell_price.border = thin_border
        for c, qty in enumerate(quantities, 3):
            cell = ws2.cell(row=r, column=c, value=qty if qty > 0 else 0)
            cell.number_format = '0.00'
            cell.border = thin_border

    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 14
    for c in range(3, 3 + len(ingredient_names)):
        from openpyxl.utils import get_column_letter
        ws2.column_dimensions[get_column_letter(c)].width = 18
    ws2.freeze_panes = "C2"

    # ===================================================================
    # Sheet3: Food Cost Analysis (EMPTY - agent must build this)
    # ===================================================================
    ws3 = wb.create_sheet("Food Cost Analysis")

    headers3 = ["Menu Item", "Selling Price", "Total Ingredient Cost", "Food Cost %", "Status"]
    for c, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=c, value=h)
    style_header(ws3, 1, len(headers3))

    # Pre-fill menu item names and selling prices (static data only)
    for r, item in enumerate(menu_items, 2):
        ws3.cell(row=r, column=1, value=item[0]).border = thin_border
        cell_price = ws3.cell(row=r, column=2, value=item[1])
        cell_price.number_format = '$#,##0.00'
        cell_price.border = thin_border
        # Columns C, D, E intentionally left empty - agent must add formulas
        for c in range(3, 6):
            ws3.cell(row=r, column=c).border = thin_border

    ws3.column_dimensions["A"].width = 26
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 16
    ws3.column_dimensions["E"].width = 16
    ws3.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
