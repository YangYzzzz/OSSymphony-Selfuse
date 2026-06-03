"""
Reward Script: Restaurant menu and ingredient cost calculator
Task ID: calc_grs_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30) - SUMPRODUCT cost formulas in Sheet3 column C (rows 2-9)
  Component 2 (0.20) - Food Cost % formulas in Sheet3 column D (rows 2-9)
  Component 3 (0.15) - Status formulas with "Review Pricing" flag in Sheet3 column E (rows 2-9)
  Component 4 (0.20) - Chart exists on Sheet3 comparing food cost percentages
  Component 5 (0.15) - Conditional formatting on Sheet3 for high-cost items
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_047'


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            import time
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: "Food Cost Analysis" sheet (or a third sheet) must exist
    sheet_names = wb.sheetnames
    # Try to find the food cost analysis sheet (Sheet3)
    fca_sheet = None
    for name in sheet_names:
        if 'cost' in name.lower() or 'analysis' in name.lower() or 'food' in name.lower():
            fca_sheet = wb[name]
            break
    if fca_sheet is None and len(sheet_names) >= 3:
        fca_sheet = wb[sheet_names[2]]
    if fca_sheet is None:
        print("FAIL: No Food Cost Analysis sheet found")
        print("REWARD: 0.0")
        return 0.0

    print(f"INFO: Using sheet '{fca_sheet.title}' as Food Cost Analysis sheet")

    # Component 1: SUMPRODUCT (or VLOOKUP-based) cost formulas in column C (0.30 points)
    # The task says "uses VLOOKUP to get ingredient costs from Sheet1" but golden uses SUMPRODUCT.
    # We accept any formula that references ingredient costs from Sheet1.
    try:
        formula_count = 0
        for row in range(2, 10):  # rows 2-9 (8 menu items)
            cell_val = fca_sheet.cell(row=row, column=3).value
            if cell_val is not None and isinstance(cell_val, str):
                cell_upper = cell_val.upper()
                # Accept SUMPRODUCT, VLOOKUP, or any formula referencing ingredient data
                if cell_upper.startswith('=') and ('INGREDIENT' in cell_upper or 'RECIPE' in cell_upper or 'VLOOKUP' in cell_upper or 'SUMPRODUCT' in cell_upper or 'SHEET1' in cell_upper):
                    formula_count += 1
        if formula_count >= 6:
            # At least 6 of 8 rows have cost formulas
            points = 0.30
            print(f"PASS: Component 1 - {formula_count}/8 rows have ingredient cost formulas in column C ({points} pts)")
            total_score += points
        elif formula_count >= 3:
            points = 0.15
            print(f"PARTIAL: Component 1 - {formula_count}/8 rows have cost formulas (partial: {points} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 1 - Only {formula_count}/8 rows have ingredient cost formulas in column C")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Food Cost % formulas in column D (0.20 points)
    # Should be =C/B or equivalent (cost divided by selling price)
    try:
        pct_formula_count = 0
        for row in range(2, 10):
            cell_val = fca_sheet.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str):
                cell_upper = cell_val.upper().replace(" ", "")
                # Accept formulas like =C2/B2 or =C$2/B$2 etc.
                if cell_upper.startswith('=') and ('C' in cell_upper and 'B' in cell_upper and '/' in cell_upper):
                    pct_formula_count += 1
                elif cell_upper.startswith('=') and '/' in cell_upper:
                    # More general: any division formula
                    pct_formula_count += 1
        if pct_formula_count >= 6:
            points = 0.20
            print(f"PASS: Component 2 - {pct_formula_count}/8 rows have Food Cost % formulas in column D ({points} pts)")
            total_score += points
        elif pct_formula_count >= 3:
            points = 0.10
            print(f"PARTIAL: Component 2 - {pct_formula_count}/8 rows have % formulas (partial: {points} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 2 - Only {pct_formula_count}/8 rows have Food Cost % formulas in column D")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Status formulas flagging items over 35% as "Review Pricing" (0.15 points)
    # Should be =IF(D>0.35,"Review Pricing","OK") or similar
    try:
        status_formula_count = 0
        for row in range(2, 10):
            cell_val = fca_sheet.cell(row=row, column=5).value
            if cell_val is not None and isinstance(cell_val, str):
                cell_upper = cell_val.upper().replace(" ", "")
                # Accept IF formulas that reference 0.35 or 35% and "REVIEW"
                if cell_upper.startswith('=') and 'IF' in cell_upper and ('0.35' in cell_upper or '35%' in cell_upper or '35' in cell_upper):
                    status_formula_count += 1
                elif cell_upper.startswith('=') and 'IF' in cell_upper and 'REVIEW' in cell_upper:
                    status_formula_count += 1
        if status_formula_count >= 6:
            points = 0.15
            print(f"PASS: Component 3 - {status_formula_count}/8 rows have status IF formulas in column E ({points} pts)")
            total_score += points
        elif status_formula_count >= 3:
            points = 0.075
            print(f"PARTIAL: Component 3 - {status_formula_count}/8 status formulas (partial: {points} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 3 - Only {status_formula_count}/8 rows have status formulas in column E")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Chart exists on the Food Cost Analysis sheet (0.20 points)
    # Task requires "a chart comparing food cost percentages across all menu items"
    try:
        charts = fca_sheet._charts
        if len(charts) >= 1:
            chart = charts[0]
            # Check it has at least 1 series (data)
            if len(chart.series) >= 1:
                points = 0.20
                print(f"PASS: Component 4 - Chart found on '{fca_sheet.title}' with {len(chart.series)} series ({points} pts)")
                total_score += points
            else:
                points = 0.10
                print(f"PARTIAL: Component 4 - Chart found but has no data series ({points} pts)")
                total_score += points
        else:
            # Also check other sheets for the chart
            chart_found = False
            for sn in sheet_names:
                ws_check = wb[sn]
                if len(ws_check._charts) > 0:
                    chart_found = True
                    ch = ws_check._charts[0]
                    if len(ch.series) >= 1:
                        points = 0.15  # slight deduction for wrong sheet
                        print(f"PARTIAL: Component 4 - Chart found on '{sn}' instead of analysis sheet ({points} pts)")
                        total_score += points
                    break
            if not chart_found:
                print(f"FAIL: Component 4 - No chart found on any sheet")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting for high-cost items (0.15 points)
    # Task requires conditional formatting for items where food cost exceeds 35%
    try:
        cf_rules = list(fca_sheet.conditional_formatting)
        if len(cf_rules) >= 1:
            # Check that at least one rule targets the food cost area
            relevant_cf = False
            for cf in cf_rules:
                cf_range = str(cf).upper()
                for rule in cf.rules:
                    # Check if rule involves 0.35 threshold or references column D
                    rule_formula = str(rule.formula) if rule.formula else ""
                    rule_operator = str(rule.operator) if rule.operator else ""
                    if ('0.35' in rule_formula or '35' in rule_formula or
                            rule_operator in ('greaterThan', 'greaterThanOrEqual') or
                            'D' in cf_range):
                        relevant_cf = True
                        break
                if relevant_cf:
                    break
            if relevant_cf:
                points = 0.15
                print(f"PASS: Component 5 - Conditional formatting with 35% threshold found ({points} pts)")
                total_score += points
            else:
                points = 0.075
                print(f"PARTIAL: Component 5 - Conditional formatting found but unclear if 35% related ({points} pts)")
                total_score += points
        else:
            print(f"FAIL: Component 5 - No conditional formatting rules found on '{fca_sheet.title}'")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
