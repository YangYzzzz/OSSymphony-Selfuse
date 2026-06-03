"""
Initial Setup: Circle Invalid Data task for LibreOffice Calc
Task ID: calc_dop_validate_circle_027
Domain: libreoffice_calc

Creates a Ratings sheet with 79 review records, data validation (whole 1-5)
already applied to column D, and 9 pre-existing invalid values in that column.

Saves as .ods (LibreOffice native format) so the golden file can persist
the Circle Invalid Data state via table:marked-invalid="true" in content.xml.
"""

import os
import subprocess
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_circle_027'
XLSX_TMP = f'{WORKDIR}/{TASK_ID}_initial_tmp.xlsx'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.ods'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Ratings'

    # Headers
    headers = ['Review ID', 'Product', 'Reviewer', 'Rating', 'Comment']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Product and reviewer data pools
    products = [
        'UltraBook Pro 15', 'SmartWatch Series 4', 'BlueAir Headphones',
        'QuickCharge 65W Adapter', 'ErgoPad Wireless Mouse', 'ClearView Monitor 27"',
        'NoiseCanceller X3', 'FlexStand Laptop Riser', 'SolidState USB Hub',
        'PixelShot Camera Kit', 'CompactRouter AC1200', 'BrightPad Tablet 10"',
        'PowerDock 6-Port', 'SwiftKey Mechanical Keyboard', 'AquaFit Smart Bottle',
        'VoiceAssist Mini', 'TrackPoint Controller', 'SleepWell Smart Pillow',
        'DeskOrganizer Pro', 'CoolBreeze USB Fan'
    ]

    reviewers = [
        'Aiden Clarke', 'Sofia Nguyen', 'Marcus Lee', 'Priya Patel', 'Daniel Kim',
        'Elena Vasquez', "James O'Brien", 'Yuki Tanaka', 'Olivia Santos', 'Ethan Brooks',
        'Maya Johnson', 'Lucas Ferreira', 'Chloe Williams', 'Noah Ahmed', 'Zoe Martinez',
        'Liam Chen', 'Amara Osei', 'Benjamin Russo', 'Hannah Park', 'Owen Campbell',
        'Fatima Hassan', 'Ryan Mueller', 'Isabelle Tremblay', 'Carlos Rivera', 'Nora Schmidt'
    ]

    comments = [
        'Great product, works exactly as described.',
        'Arrived on time, packaging was excellent.',
        'Good value for money, would recommend.',
        'Works well but setup instructions could be clearer.',
        'Exceeded my expectations, very happy.',
        'Decent quality but a bit pricey.',
        'Perfect for everyday use.',
        'Had a minor issue but customer support resolved it quickly.',
        'Solid build quality, feels premium.',
        'Easy to set up and use straight out of the box.',
        'Good performance, slight noise issue but manageable.',
        'Looks great on my desk, very functional.',
        'Exactly what I needed for remote work.',
        'Durable and reliable after several months of use.',
        'Value is outstanding for this price range.',
        'Would buy again and recommend to friends.',
        'Does the job without any fuss.',
        'Pleasantly surprised by the quality.',
        'Minor cosmetic flaw but functionally perfect.',
        'Best purchase in this category this year.',
    ]

    import random
    random.seed(42)

    # Exactly 9 invalid values in D2:D80
    # Invalid values: 2 zeros, 3 sixes, 2 negatives, 2 over-10
    invalid_values = [0, 0, 6, 6, 6, -1, -3, 11, 15]
    # Randomly select 9 row positions from rows 2-80 (indices 0-78)
    row_indices = list(range(79))
    random.shuffle(row_indices)
    invalid_rows = set(row_indices[:9])

    invalid_iter = iter(invalid_values)

    for i in range(79):
        row_num = i + 2
        review_id = f'REV-{1000 + i + 1}'
        product = products[i % len(products)]
        reviewer = reviewers[i % len(reviewers)]
        comment = comments[i % len(comments)]

        if i in invalid_rows:
            rating = next(invalid_iter)
        else:
            # Valid ratings: 1-5
            rating = random.randint(1, 5)

        ws.cell(row=row_num, column=1, value=review_id)
        ws.cell(row=row_num, column=2, value=product)
        ws.cell(row=row_num, column=3, value=reviewer)
        ws.cell(row=row_num, column=4, value=rating)
        ws.cell(row=row_num, column=5, value=comment)

    # Apply data validation: whole numbers between 1 and 5 on D2:D80
    dv = DataValidation(
        type='whole',
        operator='between',
        formula1='1',
        formula2='5',
        allow_blank=False,
        showErrorMessage=True,
        error='Rating must be a whole number between 1 and 5.',
        errorTitle='Invalid Rating',
        showInputMessage=True,
        prompt='Enter a rating from 1 to 5.',
        promptTitle='Rating'
    )
    dv.sqref = 'D2:D80'
    ws.add_data_validation(dv)

    # Save as xlsx first (temporary)
    wb.save(XLSX_TMP)
    print(f'Intermediate xlsx created: {XLSX_TMP}')

    # Convert xlsx to ods using LibreOffice headless
    # LibreOffice converts data validation to ODS format natively
    result = subprocess.run(
        ['libreoffice', '--headless', '--convert-to', 'ods',
         XLSX_TMP, '--outdir', WORKDIR],
        capture_output=True, text=True, timeout=60
    )
    if result.returncode != 0:
        print('LibreOffice conversion failed:')
        print('stdout:', result.stdout)
        print('stderr:', result.stderr)
        raise RuntimeError('LibreOffice conversion to ODS failed')

    # LibreOffice names it based on the input filename
    converted_path = f'{WORKDIR}/{TASK_ID}_initial_tmp.ods'
    if not os.path.exists(converted_path):
        print(f'Expected output at {converted_path} not found.')
        print('LibreOffice stdout:', result.stdout)
        raise FileNotFoundError(f'Converted ODS not found: {converted_path}')

    # Rename to final output name
    os.rename(converted_path, OUTPUT)
    print(f'Renamed {converted_path} -> {OUTPUT}')

    # Clean up temporary xlsx
    if os.path.exists(XLSX_TMP):
        os.remove(XLSX_TMP)

    print(f'Initial file created: {OUTPUT}')
    print('Contents:')
    print('  - Format: ODS (LibreOffice native)')
    print('  - Sheet: Ratings')
    print('  - 79 review records (rows 2-80)')
    print('  - Column D: data validation (whole numbers 1-5) applied')
    print('  - Column D: 9 invalid values (0, 0, 6, 6, 6, -1, -3, 11, 15)')
    print('  - No Circle Invalid Data markers yet')


create_initial()
