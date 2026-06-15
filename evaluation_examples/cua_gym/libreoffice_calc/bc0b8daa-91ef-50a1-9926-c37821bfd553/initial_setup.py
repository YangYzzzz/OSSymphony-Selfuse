"""
Initial Setup: Import raw experiment data, prepare workbook for statistical analysis
Task ID: calc_wf_035
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# Also create the CSV source that the task references
CSV_PATH = f'{WORKDIR}/experiment_data.csv'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_experiment_data():
    """Generate 50 measurements per group with realistic values and some outliers."""
    random.seed(42)

    group_a = []
    group_b = []
    group_c = []

    # Group A: centered around 55, range mostly 35-75, a few outliers
    for i in range(50):
        if i in [3, 17, 44]:  # outlier positions
            group_a.append(round(random.uniform(90, 100), 1))
        elif i in [11, 38]:
            group_a.append(round(random.uniform(10, 18), 1))
        else:
            group_a.append(round(random.gauss(55, 8), 1))

    # Group B: centered around 45, range mostly 28-62, a few outliers
    for i in range(50):
        if i in [7, 29]:
            group_b.append(round(random.uniform(85, 100), 1))
        elif i in [22, 41, 48]:
            group_b.append(round(random.uniform(10, 16), 1))
        else:
            group_b.append(round(random.gauss(45, 7), 1))

    # Group C: centered around 65, range mostly 48-82, a few outliers
    for i in range(50):
        if i in [5, 31, 46]:
            group_c.append(round(random.uniform(10, 20), 1))
        elif i in [19]:
            group_c.append(round(random.uniform(95, 100), 1))
        else:
            group_c.append(round(random.gauss(65, 7), 1))

    return group_a, group_b, group_c


def create_csv(group_a, group_b, group_c):
    """Write the raw CSV file that the task references."""
    lines = ["Group A,Group B,Group C"]
    for i in range(50):
        lines.append(f"{group_a[i]},{group_b[i]},{group_c[i]}")
    with open(CSV_PATH, 'w') as f:
        f.write('\n'.join(lines) + '\n')
    print(f'CSV created: {CSV_PATH}')


def create_initial():
    group_a, group_b, group_c = generate_experiment_data()

    # Also create the CSV source file
    create_csv(group_a, group_b, group_c)

    wb = openpyxl.Workbook()

    # --- Sheet 1: Raw Data (the imported CSV data) ---
    ws1 = wb.active
    ws1.title = 'Raw Data'

    # Headers
    headers = ['Group A', 'Group B', 'Group C']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Data - 50 rows per group
    for i in range(50):
        ws1.cell(row=i + 2, column=1, value=group_a[i])
        ws1.cell(row=i + 2, column=2, value=group_b[i])
        ws1.cell(row=i + 2, column=3, value=group_c[i])

    # Set column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 14

    # --- Sheet 2: Statistics (empty template - to be filled by agent) ---
    ws2 = wb.create_sheet('Statistics')

    # Just basic row labels so the agent knows what goes where
    stat_labels = [
        'Statistic', 'Group A', 'Group B', 'Group C',
    ]
    # Leave Statistics sheet completely empty - the agent needs to build it

    # --- Sheet 3: Visualization (empty - agent will create chart here) ---
    ws3 = wb.create_sheet('Visualization')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
