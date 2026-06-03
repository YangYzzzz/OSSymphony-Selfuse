"""
Reward Script: Import CSV, calculate descriptive statistics, perform IQR outlier detection,
create box-plot-style visualization, and format results as professional statistics report.
Task ID: calc_wf_035
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Statistics sheet has descriptive statistics formulas
  Component 2 (0.20): Statistics sheet has IQR outlier detection section
  Component 3 (0.20): Conditional formatting on Raw Data for outlier highlighting
  Component 4 (0.15): Visualization sheet has a chart (box plot approximation)
  Component 5 (0.10): Statistics sheet professional formatting (merged cells, borders, headers)
  Component 6 (0.10): Raw Data header styling (bold, colored fill)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_035'


def persist_app_state(domain):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Required sheets must exist
    required_sheets = ['Raw Data', 'Statistics', 'Visualization']
    for sn in required_sheets:
        if sn not in wb.sheetnames:
            print(f"CRITICAL: Missing required sheet '{sn}'")
            print("REWARD: 0.0")
            return 0.0

    ws_raw = wb['Raw Data']
    ws_stats = wb['Statistics']
    ws_viz = wb['Visualization']

    # =========================================================================
    # Component 1: Descriptive statistics formulas on Statistics sheet (0.25 pts)
    # Task requires: AVERAGE, MEDIAN, MODE, STDEV, MIN, MAX, QUARTILE for each group
    # These formulas should NOT exist on initial_env (Statistics sheet is empty)
    # =========================================================================
    try:
        stat_formulas_found = 0
        # We expect formulas referencing 'Raw Data' for statistics calculations
        # Check rows 5-13 area in columns B-D for formula presence
        formula_keywords = ['AVERAGE', 'MEDIAN', 'MODE', 'STDEV', 'MIN', 'MAX', 'QUARTILE']
        found_keywords = set()

        for row in ws_stats.iter_rows(min_row=1, max_row=ws_stats.max_row, min_col=1, max_col=4):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if isinstance(val, str) and val.startswith('='):
                    upper_val = val.upper()
                    for kw in formula_keywords:
                        if kw in upper_val:
                            found_keywords.add(kw)

        # Need at least 5 of the 7 expected formula types
        if len(found_keywords) >= 5:
            print(f"PASS: Component 1 -- Found {len(found_keywords)}/7 formula types: {found_keywords} (0.25 pts)")
            total_score += 0.25
        elif len(found_keywords) >= 3:
            partial = 0.15
            print(f"PARTIAL: Component 1 -- Found {len(found_keywords)}/7 formula types: {found_keywords} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only found {len(found_keywords)}/7 formula types: {found_keywords}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Component 2: IQR outlier detection section on Statistics sheet (0.20 pts)
    # Task requires: IQR calculation, lower/upper bounds, outlier count
    # Check for IQR-related formulas and labels
    # =========================================================================
    try:
        iqr_checks_passed = 0

        # Check for IQR-related content in Statistics sheet
        has_iqr_label = False
        has_lower_bound_label = False
        has_upper_bound_label = False
        has_outlier_count_label = False
        has_iqr_formula = False
        has_bound_formula = False
        has_countif_formula = False

        for row in ws_stats.iter_rows(min_row=1, max_row=ws_stats.max_row, min_col=1, max_col=4):
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                val = cell.value
                if val is None:
                    continue
                val_str = str(val).upper()

                # Check labels
                if 'IQR' in val_str and not val_str.startswith('='):
                    has_iqr_label = True
                if 'LOWER' in val_str and 'BOUND' in val_str:
                    has_lower_bound_label = True
                if 'UPPER' in val_str and 'BOUND' in val_str:
                    has_upper_bound_label = True
                if 'OUTLIER' in val_str and 'COUNT' in val_str:
                    has_outlier_count_label = True

                # Check formulas
                if isinstance(val, str) and val.startswith('='):
                    upper_val = val.upper()
                    # IQR formula: Q3-Q1 pattern
                    if 'B13-B12' in val or 'C13-C12' in val or 'D13-D12' in val or ('QUARTILE' in upper_val and '-' in val):
                        has_iqr_formula = True
                    # Bound formulas with 1.5*IQR
                    if '1.5' in val:
                        has_bound_formula = True
                    # COUNTIF for outlier counting
                    if 'COUNTIF' in upper_val:
                        has_countif_formula = True

        if has_iqr_label:
            iqr_checks_passed += 1
        if has_lower_bound_label or has_upper_bound_label:
            iqr_checks_passed += 1
        if has_iqr_formula or has_bound_formula:
            iqr_checks_passed += 1
        if has_countif_formula:
            iqr_checks_passed += 1

        if iqr_checks_passed >= 3:
            print(f"PASS: Component 2 -- IQR outlier detection section complete ({iqr_checks_passed}/4 checks) (0.20 pts)")
            total_score += 0.20
        elif iqr_checks_passed >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 2 -- IQR section partially complete ({iqr_checks_passed}/4 checks) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- IQR section insufficient ({iqr_checks_passed}/4 checks). "
                  f"iqr_label={has_iqr_label}, bound_labels={has_lower_bound_label or has_upper_bound_label}, "
                  f"formulas={has_iqr_formula or has_bound_formula}, countif={has_countif_formula}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Conditional formatting on Raw Data for outlier highlighting (0.20 pts)
    # Task requires: highlight outlier values in red using conditional formatting
    # Initial file has 0 conditional formatting rules
    # =========================================================================
    try:
        cf_rules = list(ws_raw.conditional_formatting)
        cf_count = len(cf_rules)

        # Count total rules across all ranges
        total_rules = sum(len(cf.rules) for cf in cf_rules)

        if total_rules >= 4:
            # At least 4 rules (e.g., 2 per column for at least 2 columns: lower + upper bound)
            print(f"PASS: Component 3 -- {total_rules} conditional formatting rules found across {cf_count} ranges (0.20 pts)")
            total_score += 0.20
        elif total_rules >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 3 -- {total_rules} conditional formatting rules found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Only {total_rules} conditional formatting rules found (expected >= 4)")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================================
    # Component 4: Visualization sheet has a chart (box plot approximation) (0.15 pts)
    # Task requires: box-plot-style visualization using a chart
    # Initial Visualization sheet has 0 charts
    # =========================================================================
    try:
        chart_count = len(ws_viz._charts)
        if chart_count >= 1:
            chart = ws_viz._charts[0]
            series_count = len(chart.series)
            if series_count >= 3:
                print(f"PASS: Component 4 -- Chart found with {series_count} series (box plot approximation) (0.15 pts)")
                total_score += 0.15
            else:
                partial = 0.08
                print(f"PARTIAL: Component 4 -- Chart found but only {series_count} series ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 4 -- No charts found on Visualization sheet")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =========================================================================
    # Component 5: Statistics sheet professional formatting (0.10 pts)
    # Task requires: formatted summary table with borders and headers
    # Check for merged title cells, borders on table cells, section headers
    # Initial Statistics sheet is completely empty
    # =========================================================================
    try:
        format_score = 0.0

        # Sub-check 5a: Merged cells exist (report title, section headers)
        merged_ranges = list(ws_stats.merged_cells.ranges)
        if len(merged_ranges) >= 2:
            format_score += 0.04
            print(f"  5a: PASS -- {len(merged_ranges)} merged cell ranges found")
        else:
            print(f"  5a: FAIL -- Only {len(merged_ranges)} merged ranges (expected >= 2)")

        # Sub-check 5b: Borders on table cells (check a few cells in the stats area)
        border_count = 0
        for row in range(5, min(21, ws_stats.max_row + 1)):
            for col in range(1, min(5, ws_stats.max_column + 1)):
                cell = ws_stats.cell(row=row, column=col)
                if isinstance(cell, MergedCell):
                    continue
                if cell.border and (cell.border.top.style or cell.border.left.style or
                                     cell.border.right.style or cell.border.bottom.style):
                    border_count += 1
        if border_count >= 8:
            format_score += 0.03
            print(f"  5b: PASS -- {border_count} cells with borders found")
        else:
            print(f"  5b: FAIL -- Only {border_count} cells with borders (expected >= 8)")

        # Sub-check 5c: Section header with fill color (Descriptive Statistics / Outlier Detection)
        header_fill_found = False
        for row in range(1, min(21, ws_stats.max_row + 1)):
            cell = ws_stats.cell(row=row, column=1)
            if isinstance(cell, MergedCell):
                continue
            try:
                fill_rgb = cell.fill.fgColor.rgb
                if fill_rgb and fill_rgb != '00000000' and cell.font.bold:
                    header_fill_found = True
                    break
            except Exception:
                pass
        if header_fill_found:
            format_score += 0.03
            print(f"  5c: PASS -- Section header with colored fill found")
        else:
            print(f"  5c: FAIL -- No colored section header found")

        if format_score > 0:
            print(f"PASS: Component 5 -- Formatting score {format_score:.2f} of 0.10")
            total_score += format_score
        else:
            print(f"FAIL: Component 5 -- No professional formatting detected")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # =========================================================================
    # Component 6: Raw Data header styling (0.10 pts)
    # Task requires: formatted report layout; golden has bold white font + blue fill on headers
    # Initial Raw Data headers are plain (no bold, no fill)
    # =========================================================================
    try:
        headers_styled = 0
        for col_letter in ['A', 'B', 'C']:
            cell = ws_raw[f'{col_letter}1']
            is_bold = cell.font.bold
            has_fill = False
            try:
                fill_rgb = cell.fill.fgColor.rgb
                if fill_rgb and fill_rgb != '00000000':
                    has_fill = True
            except Exception:
                pass

            if is_bold and has_fill:
                headers_styled += 1
            elif is_bold or has_fill:
                headers_styled += 0.5

        if headers_styled >= 2:
            print(f"PASS: Component 6 -- {headers_styled}/3 Raw Data headers properly styled (0.10 pts)")
            total_score += 0.10
        elif headers_styled >= 1:
            partial = 0.05
            print(f"PARTIAL: Component 6 -- {headers_styled}/3 headers styled ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 -- Raw Data headers not styled (bold={ws_raw['A1'].font.bold})")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification (LibreOffice may have unsaved edits)
persist_app_state("libreoffice_calc")

# Execute verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
