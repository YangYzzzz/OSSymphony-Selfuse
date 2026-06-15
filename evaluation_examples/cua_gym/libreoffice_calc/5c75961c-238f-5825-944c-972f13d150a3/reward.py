"""
Reward Script: Create pivot table showing average salary by job level with department filter
Task ID: calc_pivot_028
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): "Pivot Table" sheet exists
  Component 2 (0.20): Department filter field present
  Component 3 (0.15): Header row with JobLevel and Average of Salary labels
  Component 4 (0.30): Correct average salary values for all 5 job levels
  Component 5 (0.10): Grand Total row present
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_028'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table sheet (not HRData)."""
    for sn in wb.sheetnames:
        if sn.lower() != 'hrdata':
            return sn
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A separate pivot table sheet exists (0.25 points)
    # Initial state only has 'HRData'. The task requires creating a pivot table,
    # which should be on a new sheet.
    try:
        pivot_sheet_name = find_pivot_sheet(wb)
        if pivot_sheet_name is not None:
            print(f"PASS: Component 1 - Pivot table sheet '{pivot_sheet_name}' exists (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 - No pivot table sheet found. Only sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    if pivot_sheet_name is None:
        # No pivot sheet means nothing else to check
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb[pivot_sheet_name]

    # Component 2: Department filter field present (0.20 points)
    # The pivot table should have Department as a page/filter field.
    # In the golden file: A1="Department", B1="(All)"
    try:
        # Search for "Department" as a filter label in the first few rows
        dept_filter_found = False
        for row_idx in range(1, min(6, ws.max_row + 1)):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and str(cell_val).strip().lower() == 'department':
                # Check if there's a filter value next to it (like "(All)" or a specific dept)
                filter_val = ws.cell(row=row_idx, column=2).value
                if filter_val is not None:
                    dept_filter_found = True
                    print(f"PASS: Component 2 - Department filter found at row {row_idx}: "
                          f"'{cell_val}' with value '{filter_val}' (0.20 pts)")
                    total_score += 0.20
                    break
        if not dept_filter_found:
            print(f"FAIL: Component 2 - Department filter field not found in first rows")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Header row with JobLevel and Average of Salary labels (0.15 points)
    # The pivot should have headers indicating row field (JobLevel) and data field (Average of Salary)
    try:
        header_found = False
        for row_idx in range(1, min(10, ws.max_row + 1)):
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value
            a_str = str(a_val).strip().lower() if a_val else ''
            b_str = str(b_val).strip().lower() if b_val else ''
            if 'joblevel' in a_str.replace(' ', '').replace('_', ''):
                if 'average' in b_str and 'salary' in b_str:
                    header_found = True
                    print(f"PASS: Component 3 - Headers found at row {row_idx}: "
                          f"'{a_val}' | '{b_val}' (0.15 pts)")
                    total_score += 0.15
                    break
        if not header_found:
            print(f"FAIL: Component 3 - JobLevel/Average of Salary headers not found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Correct average salary values for all 5 job levels (0.30 points)
    # Expected: Junior=45000, Mid=65000, Senior=85000, Lead=105000, Director=130000
    # Each correct value = 0.06 points (5 x 0.06 = 0.30)
    try:
        expected_values = {
            'junior': 45000,
            'mid': 65000,
            'senior': 85000,
            'lead': 105000,
            'director': 130000,
        }
        found_count = 0
        for row_idx in range(1, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value
            b_val = ws.cell(row=row_idx, column=2).value
            if a_val and str(a_val).strip().lower() in expected_values:
                level_key = str(a_val).strip().lower()
                expected = expected_values[level_key]
                if b_val is not None:
                    try:
                        actual = float(b_val)
                        if abs(actual - expected) <= 1.0:
                            found_count += 1
                            print(f"  PASS: {level_key} salary = {actual} (expected {expected})")
                        else:
                            print(f"  FAIL: {level_key} salary = {actual} (expected {expected})")
                    except (ValueError, TypeError):
                        print(f"  FAIL: {level_key} salary value not numeric: {b_val}")
                else:
                    print(f"  FAIL: {level_key} salary is None")

        points = found_count * 0.06
        if found_count == 5:
            print(f"PASS: Component 4 - All 5 job level averages correct (0.30 pts)")
            total_score += 0.30
        elif found_count > 0:
            print(f"PARTIAL: Component 4 - {found_count}/5 job level averages correct ({points:.2f} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 4 - No correct job level averages found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Grand Total row present (0.10 points)
    # The pivot should have a grand total/summary row
    try:
        grand_total_found = False
        for row_idx in range(1, ws.max_row + 1):
            a_val = ws.cell(row=row_idx, column=1).value
            if a_val and 'grand total' in str(a_val).strip().lower():
                b_val = ws.cell(row=row_idx, column=2).value
                if b_val is not None:
                    grand_total_found = True
                    print(f"PASS: Component 5 - Grand Total row found at row {row_idx} "
                          f"with value {b_val} (0.10 pts)")
                    total_score += 0.10
                    break
        if not grand_total_found:
            print(f"FAIL: Component 5 - Grand Total row not found")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
