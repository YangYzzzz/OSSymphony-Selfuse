"""
Initial Setup: Create HR data spreadsheet for pivot table task
Task ID: calc_pivot_028
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HRData'

    # --- Headers ---
    headers = ['EmpID', 'Name', 'Department', 'JobLevel', 'Salary', 'StartDate']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        bottom=Side(style="thin", color="000000")
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Departments and Job Levels ---
    departments = ['Engineering', 'Marketing', 'Finance', 'Operations', 'Human Resources']
    job_levels = ['Junior', 'Mid', 'Senior', 'Lead', 'Director']

    # Target average salaries per job level (ground truth)
    target_avg = {
        'Junior': 45000,
        'Mid': 65000,
        'Senior': 85000,
        'Lead': 105000,
        'Director': 130000,
    }

    # Distribution: 150 employees across 5 levels
    # 30 per level to keep averages exact
    level_counts = {'Junior': 30, 'Mid': 30, 'Senior': 30, 'Lead': 30, 'Director': 30}

    # Generate salaries for each level that average exactly to the target
    def generate_salaries(target, count):
        """Generate count salaries that average exactly to target."""
        salaries = []
        spread = int(target * 0.15)  # 15% spread
        for i in range(count - 1):
            s = target + random.randint(-spread, spread)
            # Round to nearest 500
            s = round(s / 500) * 500
            salaries.append(s)
        # Last salary adjusts to make average exact
        current_sum = sum(salaries)
        last_salary = target * count - current_sum
        salaries.append(last_salary)
        random.shuffle(salaries)
        return salaries

    # First/last names for realistic data
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Maria', 'Chen',
        'Robert', 'Aisha', 'Thomas', 'Yuki', 'Michael', 'Fatima', 'William',
        'Sofia', 'Daniel', 'Amara', 'Alexander', 'Lin', 'Christopher', 'Nadia',
        'Andrew', 'Zara', 'Benjamin', 'Keiko', 'Joshua', 'Ines', 'Ryan', 'Mei',
        'Ethan', 'Layla', 'Nathan', 'Olga', 'Patrick', 'Rashida', 'Samuel', 'Tanya',
        'Kevin', 'Uma', 'Brian', 'Vera', 'Scott', 'Wendy', 'Adam', 'Xena',
        'Lucas', 'Diana', 'Gabriel', 'Hannah', 'Oscar', 'Iris', 'Felix', 'Julia',
        'Victor', 'Karen', 'George', 'Lily', 'Henry', 'Monica', 'Ian', 'Nina',
        'Jack', 'Paula', 'Keith', 'Rachel', 'Leo', 'Sandra', 'Max', 'Tina',
        'Neil', 'Ursula', 'Owen', 'Vivian', 'Peter', 'Wanda', 'Quinn', 'Yvette',
        'Ray', 'Alice', 'Sean', 'Beth', 'Troy', 'Carol', 'Vincent', 'Debra',
        'Walter', 'Eva', 'Xavier', 'Fiona', 'Yusuf', 'Gloria', 'Zachary', 'Holly',
        'Aaron', 'Ingrid', 'Barry', 'Janet', 'Carl', 'Kim', 'Derek', 'Laura',
        'Eric', 'Megan', 'Frank', 'Naomi', 'Grant', 'Olivia', 'Hugo', 'Pam',
        'Ivan', 'Rosa', 'Jake', 'Sylvia', 'Kurt', 'Tracy', 'Luke', 'Valentina',
        'Mark', 'Whitney', 'Noel', 'Ximena', 'Otto', 'Yolanda', 'Paul', 'Zoe',
        'Rick', 'April', 'Stan', 'Brenda', 'Tim', 'Chloe', 'Umar', 'Dawn',
        'Vince', 'Elaine', 'Wade', 'Frances', 'Xander', 'Gina', 'Yuri', 'Heidi',
        'Zack', 'Isabel', 'Amos', 'Joy', 'Blake', 'Kara',
    ]
    last_names = [
        'Chen', 'Johnson', 'Kowalski', 'Rivera', 'Nakamura', 'Okonkwo', 'Mueller',
        'Santos', 'Kim', 'Patel', 'Andersen', 'Bianchi', 'Garcia', 'Tanaka',
        'Williams', 'Johansson', 'Fernandez', 'Nguyen', 'Park', 'O\'Brien',
        'Schneider', 'Ivanova', 'Martinez', 'Suzuki', 'Brown', 'Lindqvist',
        'Hernandez', 'Watanabe', 'Davis', 'Petrov', 'Lopez', 'Yamamoto',
        'Wilson', 'Novak', 'Gonzalez', 'Sato', 'Moore', 'Horvat', 'Rodriguez',
        'Takahashi', 'Taylor', 'Kovacs', 'Thomas', 'Ito', 'Anderson', 'Berg',
        'Jackson', 'Mori', 'White', 'Fischer',
    ]

    # Start dates
    years = [2019, 2020, 2021, 2022, 2023, 2024, 2025]
    months = list(range(1, 13))
    days_in_month = {1: 28, 2: 28, 3: 28, 4: 28, 5: 28, 6: 28,
                     7: 28, 8: 28, 9: 28, 10: 28, 11: 28, 12: 28}

    # Build all employee records
    employees = []
    emp_id = 1
    for level in job_levels:
        count = level_counts[level]
        salaries = generate_salaries(target_avg[level], count)
        for i in range(count):
            dept = departments[i % len(departments)]
            fn = first_names[(emp_id - 1) % len(first_names)]
            ln = last_names[(emp_id - 1) % len(last_names)]
            name = f'{fn} {ln}'
            year = random.choice(years)
            month = random.randint(1, 12)
            day = random.randint(1, 28)
            start_date = f'{year:04d}-{month:02d}-{day:02d}'
            employees.append([emp_id, name, dept, level, salaries[i], start_date])
            emp_id += 1

    # Shuffle employees so they're not grouped by level
    random.shuffle(employees)

    # Re-assign sequential EmpIDs after shuffle
    for idx, emp in enumerate(employees):
        emp[0] = idx + 1

    # Write data
    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Number format for salary column
    for r in range(2, 152):
        ws.cell(row=r, column=5).number_format = '#,##0'

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify averages
    from collections import defaultdict
    level_sums = defaultdict(float)
    level_cnts = defaultdict(int)
    for emp in employees:
        level_sums[emp[3]] += emp[4]
        level_cnts[emp[3]] += 1
    for level in job_levels:
        avg = level_sums[level] / level_cnts[level]
        print(f'  {level}: avg={avg:.2f} (target={target_avg[level]})')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
