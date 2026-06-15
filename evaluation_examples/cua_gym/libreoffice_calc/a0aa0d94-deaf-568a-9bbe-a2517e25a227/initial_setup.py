"""
Initial Setup: Product sales spreadsheet with 8 products, pre-task state
Task ID: osworld_calc_multi_chart_computed_011
Domain: libreoffice_calc

Creates a product sales spreadsheet with product names and total sales values.
Column C header exists but contains no formulas (agent must add them).
No charts are present in the initial state.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Product Sales ---
    ws = wb.active
    ws.title = "Product Sales"

    # Headers
    ws["A1"] = "Product"
    ws["B1"] = "Total Sales (USD)"
    ws["C1"] = "Contribution (%)"

    # 8 products with realistic sales data
    # Column C intentionally left empty (no formulas — agent must add them)
    products = [
        ("Alpine Outdoor Gear",     142580),
        ("BlueSky Wireless Earbuds", 98340),
        ("ClearView Monitor Pro",   215760),
        ("DeltaForm Running Shoes",  87920),
        ("EcoBreeze Air Purifier",  163450),
        ("FlexCore Yoga Mat",        54270),
        ("GardenPro Tool Set",       76890),
        ("HorizonX Smart Watch",    193640),
    ]

    for r, (product, sales) in enumerate(products, 2):
        ws.cell(row=r, column=1, value=product)
        ws.cell(row=r, column=2, value=sales)
        # Column C intentionally left blank

    # Set column widths for readability
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
