"""
Reward Script: Add percentage contribution column and two charts to product sales spreadsheet
Task ID: osworld_calc_multi_chart_computed_011
Domain: libreoffice_calc
Scoring:
  - Component 1: Percentage contribution formulas in C2:C9 (0.4 points)
  - Component 2: Bar chart titled 'Product Sales Totals' using column B data (0.3 points)
  - Component 3: Line chart titled 'Sales Contribution (%)' using column C data (0.3 points)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_011'


def get_chart_title_text(chart):
    """Extract plain text from a chart title object."""
    try:
        # Standard openpyxl title path
        return chart.title.tx.rich.p[0].r[0].t
    except Exception:
        pass
    try:
        # Fallback: str representation may contain title
        return str(chart.title)
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Add percentage contribution column (C2:C9 with formulas = Bn/SUM(B2:B9)*100)
          and two charts:
          1. Bar chart titled 'Product Sales Totals' referencing column B
          2. Line chart titled 'Sales Contribution (%)' referencing column C
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: The 'Product Sales' sheet must exist
    if 'Product Sales' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Product Sales' not found. Sheets:", wb.sheetnames)
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Product Sales']

    # -------------------------------------------------------------------------
    # Component 1: Percentage contribution formulas in C2:C9 (0.4 points)
    #
    # Each cell must contain a formula that references the corresponding Bn cell
    # divided by a SUM of B2:B9 (or equivalent), multiplied by 100.
    # We check that:
    #   (a) All 8 cells in C2:C9 are non-empty (have a formula, not None)
    #   (b) Each formula contains the pattern Bn/... or .../Bn (references row-aligned B)
    #   (c) Each formula references SUM of B column
    #   (d) Multiplied by 100 (or equivalent)
    # Partial credit: 0.2 pts if formulas present but pattern imperfect
    # -------------------------------------------------------------------------
    try:
        c_cells = [ws.cell(row=r, column=3).value for r in range(2, 10)]
        non_empty = sum(1 for v in c_cells if v is not None)

        if non_empty == 0:
            print(f"FAIL: Component 1 — Column C (C2:C9) has no values (all None)")
        elif non_empty < 8:
            print(f"FAIL: Component 1 — Only {non_empty}/8 cells in C2:C9 have values")
        else:
            # All 8 cells have values; check formula quality
            formula_quality = 0
            for i, val in enumerate(c_cells):
                row_num = i + 2
                val_str = str(val).upper().replace(" ", "")
                # Must reference the B column cell for this row (B2, B3, ... B9)
                has_b_ref = f"B{row_num}" in val_str
                # Must reference a SUM function on B column
                has_sum = "SUM" in val_str and "B" in val_str
                # Must multiply by 100 or divide by a fraction (check for *100 or /0.01)
                has_pct = "*100" in val_str or "/0.01" in val_str
                if has_b_ref and has_sum and has_pct:
                    formula_quality += 1

            if formula_quality == 8:
                print(f"PASS: Component 1 — All 8 percentage formulas present and correct "
                      f"(e.g., C2={c_cells[0]}) (0.4 pts)")
                total_score += 0.4
            elif formula_quality >= 4:
                print(f"PARTIAL: Component 1 — {formula_quality}/8 cells have full formula "
                      f"quality; awarding 0.2 pts")
                total_score += 0.2
            else:
                print(f"FAIL: Component 1 — Only {formula_quality}/8 cells pass full formula "
                      f"check. Sample C2={c_cells[0]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Bar chart titled 'Product Sales Totals' using column B (0.3 points)
    #
    # Must find a BarChart (type='col' or any BarChart) with:
    #   - Title containing 'Product Sales Totals'
    #   - At least 1 series whose val references column B data
    # -------------------------------------------------------------------------
    try:
        charts = ws._charts
        if not charts:
            print(f"FAIL: Component 2 — No charts found in 'Product Sales' sheet")
        else:
            bar_chart_found = False
            for chart in charts:
                # Check if it's a bar chart
                if type(chart).__name__ != 'BarChart':
                    continue
                # Check title
                title_text = get_chart_title_text(chart)
                title_matches = (
                    title_text is not None and
                    'product sales totals' in title_text.lower()
                )
                if not title_matches:
                    continue
                # Check that at least one series references column B
                series_refs_b = False
                for ser in chart.series:
                    try:
                        num_ref_f = ser.val.numRef.f if ser.val and ser.val.numRef else ""
                        # Reference should include column B data range (e.g. $B$2:$B$9)
                        if "$B$" in num_ref_f or "'Product Sales'!B" in num_ref_f:
                            series_refs_b = True
                            break
                    except Exception:
                        pass
                if title_matches and series_refs_b:
                    print(f"PASS: Component 2 — Bar chart 'Product Sales Totals' found with "
                          f"column B data reference (0.3 pts)")
                    total_score += 0.3
                    bar_chart_found = True
                    break
                elif title_matches:
                    # Title matches but series doesn't reference B
                    print(f"FAIL: Component 2 — Bar chart title matches but series does not "
                          f"reference column B data")
                    break

            if not bar_chart_found:
                # Report what charts exist for debugging
                chart_summary = [
                    (type(c).__name__, get_chart_title_text(c)) for c in charts
                ]
                print(f"FAIL: Component 2 — No bar chart titled 'Product Sales Totals' with "
                      f"column B data found. Charts present: {chart_summary}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Line chart titled 'Sales Contribution (%)' using column C (0.3 points)
    #
    # Must find a LineChart with:
    #   - Title containing 'Sales Contribution (%)'
    #   - At least 1 series whose val references column C data
    # -------------------------------------------------------------------------
    try:
        charts = ws._charts
        if not charts:
            print(f"FAIL: Component 3 — No charts found in 'Product Sales' sheet")
        else:
            line_chart_found = False
            for chart in charts:
                # Check if it's a line chart
                if type(chart).__name__ != 'LineChart':
                    continue
                # Check title
                title_text = get_chart_title_text(chart)
                title_matches = (
                    title_text is not None and
                    'sales contribution' in title_text.lower() and
                    '%' in title_text
                )
                if not title_matches:
                    continue
                # Check that at least one series references column C
                series_refs_c = False
                for ser in chart.series:
                    try:
                        num_ref_f = ser.val.numRef.f if ser.val and ser.val.numRef else ""
                        # Reference should include column C data range (e.g. $C$2:$C$9)
                        if "$C$" in num_ref_f or "'Product Sales'!C" in num_ref_f:
                            series_refs_c = True
                            break
                    except Exception:
                        pass
                if title_matches and series_refs_c:
                    print(f"PASS: Component 3 — Line chart 'Sales Contribution (%)' found with "
                          f"column C data reference (0.3 pts)")
                    total_score += 0.3
                    line_chart_found = True
                    break
                elif title_matches:
                    print(f"FAIL: Component 3 — Line chart title matches but series does not "
                          f"reference column C data")
                    break

            if not line_chart_found:
                chart_summary = [
                    (type(c).__name__, get_chart_title_text(c)) for c in charts
                ]
                print(f"FAIL: Component 3 — No line chart titled 'Sales Contribution (%)' with "
                      f"column C data found. Charts present: {chart_summary}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
