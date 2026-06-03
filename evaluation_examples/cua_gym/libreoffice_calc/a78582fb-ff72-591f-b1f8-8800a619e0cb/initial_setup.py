"""
Initial Setup: Customer satisfaction survey results dashboard
Task ID: calc_wf_072
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# ---------- Deterministic seed for reproducibility ----------
random.seed(42)

# ---------- Survey Data Generation ----------
AGE_GROUPS = ['18-24', '25-34', '35-44', '45-54', '55-64', '65+']
REGIONS = ['North', 'South', 'East', 'West']
PRODUCTS = ['CloudSync Pro', 'DataVault', 'SecureNet', 'AnalytiX', 'DevFlow']

FIRST_NAMES = [
    'Sarah', 'Marcus', 'Priya', 'James', 'Elena', 'Wei', 'Aisha', 'Carlos',
    'Yuki', 'Olga', 'Raj', 'Fatima', 'Tomasz', 'Lena', 'Hassan', 'Maya',
    'Dmitri', 'Amara', 'Chen', 'Sofia', 'Kwame', 'Ingrid', 'Ali', 'Rosa',
    'Viktor', 'Nadia', 'Kofi', 'Hana', 'Liam', 'Zara'
]

POSITIVE_COMMENTS = [
    'Great product, very intuitive',
    'Exceeded my expectations',
    'Solid performance and reliable',
    'Customer support was excellent',
    'Easy to set up and configure',
    'Very satisfied with the purchase',
    'Would definitely recommend',
    'Good value for the price',
    'Streamlined our workflow significantly',
    'The update made a huge difference',
]

NEUTRAL_COMMENTS = [
    'Decent product, nothing special',
    'Works as expected',
    'Average experience overall',
    'Could use some improvements',
    'Meets basic requirements',
    'Some features are unnecessary',
    'Documentation could be better',
    'Acceptable performance',
]

NEGATIVE_COMMENTS = [
    'Too slow for our needs',
    'Interface is confusing',
    'Missing key features',
    'Not worth the premium price',
    'Frequent crashes and bugs',
    'Poor onboarding experience',
]

def generate_comment(satisfaction, recommend):
    if satisfaction >= 4 and recommend >= 7:
        return random.choice(POSITIVE_COMMENTS)
    elif satisfaction <= 2 or recommend <= 4:
        return random.choice(NEGATIVE_COMMENTS)
    else:
        return random.choice(NEUTRAL_COMMENTS)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Survey Data ---
    ws1 = wb.active
    ws1.title = 'Survey Data'

    headers = ['Date', 'Age Group', 'Region', 'Product', 'Satisfaction', 'Effort', 'Recommend', 'Comments']
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Generate 200 survey responses across 4 quarters of 2025
    q_start_dates = [
        datetime(2025, 1, 1),   # Q1
        datetime(2025, 4, 1),   # Q2
        datetime(2025, 7, 1),   # Q3
        datetime(2025, 10, 1),  # Q4
    ]
    q_end_dates = [
        datetime(2025, 3, 31),
        datetime(2025, 6, 30),
        datetime(2025, 9, 30),
        datetime(2025, 12, 31),
    ]

    for i in range(200):
        row = i + 2
        # Distribute roughly evenly across quarters
        q = i % 4
        start = q_start_dates[q]
        end = q_end_dates[q]
        delta = (end - start).days
        survey_date = start + timedelta(days=random.randint(0, delta))

        age_group = random.choice(AGE_GROUPS)
        region = random.choice(REGIONS)
        product = random.choice(PRODUCTS)
        satisfaction = random.choices([1, 2, 3, 4, 5], weights=[5, 10, 20, 35, 30])[0]
        effort = random.choices([1, 2, 3, 4, 5], weights=[25, 30, 25, 12, 8])[0]
        # NPS: recommend 0-10
        recommend = random.choices(
            list(range(11)),
            weights=[2, 2, 3, 4, 5, 8, 10, 15, 18, 18, 15]
        )[0]
        comment = generate_comment(satisfaction, recommend)

        ws1.cell(row=row, column=1, value=survey_date).number_format = 'yyyy-mm-dd'
        ws1.cell(row=row, column=2, value=age_group)
        ws1.cell(row=row, column=3, value=region)
        ws1.cell(row=row, column=4, value=product)
        ws1.cell(row=row, column=5, value=satisfaction)
        ws1.cell(row=row, column=6, value=effort)
        ws1.cell(row=row, column=7, value=recommend)
        ws1.cell(row=row, column=8, value=comment)

    # Set column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 10
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 40

    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Dashboard (empty template, NO formulas or charts) ---
    ws2 = wb.create_sheet('Dashboard')

    # Title
    ws2.merge_cells('A1:H1')
    ws2['A1'] = 'Customer Satisfaction Survey Dashboard'
    ws2['A1'].font = Font(size=18, bold=True, color="2F5496")
    ws2['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 40

    # KPI Section headers (row 3)
    ws2['A3'] = 'Key Performance Indicators'
    ws2['A3'].font = Font(size=14, bold=True)

    kpi_labels = ['NPS Score', 'CSAT Score', 'CES Score']
    for col, label in enumerate(kpi_labels, 1):
        cell = ws2.cell(row=4, column=col * 2 - 1, value=label)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center")
        # Value cell placeholder (next to label) -- left EMPTY for the task
        val_cell = ws2.cell(row=5, column=col * 2 - 1)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.font = Font(size=16)

    # Demographic Breakdown section header (row 8)
    ws2['A8'] = 'Demographic Breakdown'
    ws2['A8'].font = Font(size=14, bold=True)

    # Crosstab headers - Age Group x Region (labels only, no data)
    ws2['A9'] = 'Age Group \\ Region'
    ws2['A9'].font = Font(bold=True)
    for col, region in enumerate(REGIONS, 2):
        cell = ws2.cell(row=9, column=col, value=region)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row, ag in enumerate(AGE_GROUPS, 10):
        ws2.cell(row=row, column=1, value=ag).font = Font(bold=True)
        # Data cells left EMPTY

    # Quarterly Trends section header (row 18)
    ws2['A18'] = 'Quarterly Trends'
    ws2['A18'].font = Font(size=14, bold=True)

    ws2['A19'] = 'Quarter'
    ws2['A19'].font = Font(bold=True)
    ws2['B19'] = 'Avg Satisfaction'
    ws2['B19'].font = Font(bold=True)
    ws2['C19'] = 'Avg Effort'
    ws2['C19'].font = Font(bold=True)
    ws2['D19'] = 'Avg Recommend'
    ws2['D19'].font = Font(bold=True)

    quarters = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025']
    for row, q in enumerate(quarters, 20):
        ws2.cell(row=row, column=1, value=q).font = Font(bold=True)
        # Values left EMPTY

    # CSAT by Product section header (row 26)
    ws2['A26'] = 'CSAT by Product'
    ws2['A26'].font = Font(size=14, bold=True)

    ws2['A27'] = 'Product'
    ws2['A27'].font = Font(bold=True)
    ws2['B27'] = 'CSAT %'
    ws2['B27'].font = Font(bold=True)

    for row, prod in enumerate(PRODUCTS, 28):
        ws2.cell(row=row, column=1, value=prod).font = Font(bold=True)
        # Values left EMPTY

    # Column widths for Dashboard
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 14
    ws2.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
