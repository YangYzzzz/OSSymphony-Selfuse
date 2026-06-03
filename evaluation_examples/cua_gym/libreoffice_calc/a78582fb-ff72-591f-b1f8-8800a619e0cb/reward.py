"""
Reward Script: Customer Satisfaction Survey Results Dashboard
Task ID: calc_wf_072
Domain: libreoffice_calc
Scoring:
  Component 1: KPI formulas (NPS, CSAT, CES) in Dashboard — 0.20 points
  Component 2: Demographic breakdown COUNTIFS formulas — 0.20 points
  Component 3: Quarterly trend AVERAGEIFS formulas — 0.20 points
  Component 4: CSAT by Product formulas + NPS distribution data — 0.15 points
  Component 5: Charts (Pie/Donut, Bar, Line) — 0.25 points
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_072'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice state via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def has_formula(value, pattern_keyword):
    """Check if a cell value is a formula string containing the given keyword."""
    if not isinstance(value, str):
        return False
    return value.startswith('=') and pattern_keyword.upper() in value.upper()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print("CRITICAL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # =========================================================================
    # Component 1: KPI Formulas — NPS, CSAT, CES (0.20 points)
    # Golden has: A5=NPS formula (COUNTIF-based), C5=CSAT formula, E5=CES formula (AVERAGE)
    # Initial has: these cells are empty
    # =========================================================================
    try:
        kpi_score = 0.0
        a5 = ws['A5'].value
        c5 = ws['C5'].value
        e5 = ws['E5'].value

        # NPS formula: should reference COUNTIF on Recommend column (G) with >=9 and <=6
        if isinstance(a5, str) and a5.startswith('=') and 'COUNTIF' in a5.upper():
            kpi_score += 1.0 / 3.0
            print(f"PASS: A5 has NPS formula with COUNTIF")
        else:
            print(f"FAIL: A5 expected NPS formula with COUNTIF, found: {repr(a5)}")

        # CSAT formula: should reference COUNTIF on Satisfaction column (E) with >=4
        if isinstance(c5, str) and c5.startswith('=') and 'COUNTIF' in c5.upper():
            kpi_score += 1.0 / 3.0
            print(f"PASS: C5 has CSAT formula with COUNTIF")
        else:
            print(f"FAIL: C5 expected CSAT formula with COUNTIF, found: {repr(c5)}")

        # CES formula: should be AVERAGE of effort column (F)
        if isinstance(e5, str) and e5.startswith('=') and 'AVERAGE' in e5.upper():
            kpi_score += 1.0 / 3.0
            print(f"PASS: E5 has CES formula with AVERAGE")
        else:
            print(f"FAIL: E5 expected CES formula with AVERAGE, found: {repr(e5)}")

        component_1 = round(0.20 * kpi_score, 4)
        total_score += component_1
        print(f"Component 1 subtotal: {component_1:.4f}/0.20")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Demographic Breakdown — COUNTIFS formulas (0.20 points)
    # Golden has: B10:E15 with COUNTIFS formulas cross-tabbing Age Group x Region
    # Initial has: these cells are empty (only labels in A10:A15)
    # =========================================================================
    try:
        countifs_found = 0
        total_cells = 24  # 6 age groups x 4 regions = 24 cells

        for row in range(10, 16):  # rows 10-15
            for col_letter in ['B', 'C', 'D', 'E']:
                cell_val = ws[f'{col_letter}{row}'].value
                if isinstance(cell_val, str) and cell_val.startswith('=') and 'COUNTIFS' in cell_val.upper():
                    countifs_found += 1

        if countifs_found >= 20:
            # At least 20 of 24 cells have COUNTIFS — full credit
            component_2 = 0.20
            print(f"PASS: Demographic COUNTIFS — {countifs_found}/{total_cells} cells have COUNTIFS formulas (0.20 pts)")
        elif countifs_found >= 12:
            component_2 = 0.15
            print(f"PARTIAL: Demographic COUNTIFS — {countifs_found}/{total_cells} cells (0.15 pts)")
        elif countifs_found >= 4:
            component_2 = 0.10
            print(f"PARTIAL: Demographic COUNTIFS — {countifs_found}/{total_cells} cells (0.10 pts)")
        else:
            component_2 = 0.0
            print(f"FAIL: Demographic COUNTIFS — only {countifs_found}/{total_cells} cells have formulas")

        total_score += component_2
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Quarterly Trends — AVERAGEIFS formulas (0.20 points)
    # Golden has: B20:D23 with AVERAGEIFS formulas for 4 quarters x 3 metrics
    # Initial has: these cells are empty (only labels in A20:A23)
    # =========================================================================
    try:
        averageifs_found = 0
        total_quarterly_cells = 12  # 4 quarters x 3 metrics

        for row in range(20, 24):  # rows 20-23
            for col_letter in ['B', 'C', 'D']:
                cell_val = ws[f'{col_letter}{row}'].value
                if isinstance(cell_val, str) and cell_val.startswith('=') and 'AVERAGEIF' in cell_val.upper():
                    averageifs_found += 1

        if averageifs_found >= 10:
            component_3 = 0.20
            print(f"PASS: Quarterly AVERAGEIFS — {averageifs_found}/{total_quarterly_cells} cells (0.20 pts)")
        elif averageifs_found >= 6:
            component_3 = 0.15
            print(f"PARTIAL: Quarterly AVERAGEIFS — {averageifs_found}/{total_quarterly_cells} cells (0.15 pts)")
        elif averageifs_found >= 3:
            component_3 = 0.10
            print(f"PARTIAL: Quarterly AVERAGEIFS — {averageifs_found}/{total_quarterly_cells} cells (0.10 pts)")
        else:
            component_3 = 0.0
            print(f"FAIL: Quarterly AVERAGEIFS — only {averageifs_found}/{total_quarterly_cells} cells")

        total_score += component_3
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: CSAT by Product formulas + NPS distribution data (0.15 points)
    # Golden has: B28:B32 with CSAT formulas per product, F36:G38 with NPS counts
    # Initial has: these cells are empty
    # =========================================================================
    try:
        component_4 = 0.0

        # Sub-check 4a: CSAT by Product formulas (0.09 points)
        csat_product_count = 0
        for row in range(28, 33):  # rows 28-32
            cell_val = ws[f'B{row}'].value
            if isinstance(cell_val, str) and cell_val.startswith('=') and 'COUNTIF' in cell_val.upper():
                csat_product_count += 1

        if csat_product_count >= 4:
            component_4 += 0.09
            print(f"PASS: CSAT by Product — {csat_product_count}/5 formulas found (0.09 pts)")
        elif csat_product_count >= 2:
            component_4 += 0.05
            print(f"PARTIAL: CSAT by Product — {csat_product_count}/5 formulas found (0.05 pts)")
        else:
            print(f"FAIL: CSAT by Product — only {csat_product_count}/5 formulas found")

        # Sub-check 4b: NPS distribution data (0.06 points)
        # Check for Promoters/Passives/Detractors counts using COUNTIF formulas
        nps_dist_count = 0
        # Search in a wider area since NPS distribution may be in different cells
        for row in range(35, 45):
            for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                try:
                    cell_val = ws[f'{col_letter}{row}'].value
                    if isinstance(cell_val, str) and cell_val.startswith('=') and 'COUNTIF' in cell_val.upper():
                        nps_dist_count += 1
                except:
                    pass

        if nps_dist_count >= 2:
            component_4 += 0.06
            print(f"PASS: NPS distribution — {nps_dist_count} COUNTIF formulas found (0.06 pts)")
        else:
            print(f"FAIL: NPS distribution — only {nps_dist_count} COUNTIF formulas found")

        total_score += component_4
        print(f"Component 4 subtotal: {component_4:.4f}/0.15")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Charts — Pie/Donut, Bar, Line (0.25 points)
    # Golden has: 3 charts (PieChart for NPS, BarChart for CSAT, LineChart for trends)
    # Initial has: 0 charts
    # =========================================================================
    try:
        charts = ws._charts
        num_charts = len(charts)
        component_5 = 0.0

        if num_charts == 0:
            print(f"FAIL: No charts found in Dashboard (expected 3)")
        else:
            # Check for chart types
            chart_types = set()
            for chart in charts:
                type_name = type(chart).__name__
                chart_types.add(type_name)

            has_pie_or_donut = 'PieChart' in chart_types or 'DoughnutChart' in chart_types
            has_bar = 'BarChart' in chart_types or 'BarChart3D' in chart_types
            has_line = 'LineChart' in chart_types or 'LineChart3D' in chart_types

            if has_pie_or_donut:
                component_5 += 0.08
                print(f"PASS: Pie/Donut chart found for NPS gauge (0.08 pts)")
            else:
                print(f"FAIL: No Pie/Donut chart found (expected NPS gauge)")

            if has_bar:
                component_5 += 0.09
                print(f"PASS: Bar chart found for CSAT by product (0.09 pts)")
            else:
                print(f"FAIL: No Bar chart found (expected CSAT by category)")

            if has_line:
                component_5 += 0.08
                print(f"PASS: Line chart found for quarterly trends (0.08 pts)")
            else:
                print(f"FAIL: No Line chart found (expected quarterly trends)")

        total_score += component_5
        print(f"Component 5 subtotal: {component_5:.4f}/0.25")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
