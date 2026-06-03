"""
Reward Script: Extract domain name from email addresses using MID/FIND formulas
Task ID: calc_fma_mid_find_056
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): All 11 cells B2:B12 contain non-empty formula strings
  - Component 2 (0.3): Formulas use MID and FIND functions referencing column A
  - Component 3 (0.2): Formula row references are correct (each cell references its own row)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_mid_find_056'

# Expected email -> domain mappings for spot-check validation
EXPECTED_DOMAINS_APPROX = [
    'gmail', 'yahoo', 'outlook', 'company', 'startup',
    'enterprise', 'webmail', 'proton', 'icloud', 'fastmail', 'zoho'
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires:
    1. Cells B2:B12 (11 cells) must all contain formulas (were None in initial file)
    2. The formulas must use MID and FIND functions to extract domain from column A
    3. Each formula must reference the correct row number in column A
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Emails' sheet must exist
    if 'Emails' not in wb.sheetnames:
        print("FAIL: 'Emails' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Emails']

    # Component 1: All 11 cells B2:B12 contain non-empty values/formulas (0.5 points)
    # In initial file, B2:B12 are all None — so any non-None value indicates task action
    try:
        filled_cells = 0
        formula_cells = []
        for row in range(2, 13):  # rows 2-12 inclusive (11 rows)
            cell_val = ws.cell(row=row, column=2).value
            if cell_val is not None and str(cell_val).strip() != '':
                filled_cells += 1
                formula_cells.append((row, str(cell_val)))
            else:
                print(f"FAIL: Component 1 — B{row} is empty (None or blank)")

        if filled_cells == 11:
            print(f"PASS: Component 1 — All 11 cells B2:B12 contain values/formulas ({filled_cells}/11)")
            total_score += 0.5
        elif filled_cells > 0:
            # Partial credit for partially filled range
            partial = round(0.5 * (filled_cells / 11), 3)
            print(f"PARTIAL: Component 1 — {filled_cells}/11 cells B2:B12 contain values (partial: {partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 — No cells in B2:B12 contain any values")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        formula_cells = []

    # Component 2: Formulas use MID and FIND functions referencing column A (0.3 points)
    # The task requires formulas, not static text values
    # A valid formula must: start with '=', contain 'MID', contain 'FIND', reference 'A'
    try:
        if not formula_cells:
            print("FAIL: Component 2 — No formula cells found (skipping formula check)")
        else:
            valid_formula_count = 0
            for row, val in formula_cells:
                val_upper = val.upper().replace(' ', '')
                # Must be a formula (starts with '=')
                is_formula = val.startswith('=')
                # Must use MID function
                uses_mid = 'MID(' in val_upper
                # Must use FIND function
                uses_find = 'FIND(' in val_upper
                # Must reference column A (for extracting domain from email)
                refs_col_a = bool(re.search(r'A\d+', val_upper))

                if is_formula and uses_mid and uses_find and refs_col_a:
                    valid_formula_count += 1
                else:
                    reasons = []
                    if not is_formula:
                        reasons.append("not a formula")
                    if not uses_mid:
                        reasons.append("missing MID function")
                    if not uses_find:
                        reasons.append("missing FIND function")
                    if not refs_col_a:
                        reasons.append("does not reference column A")
                    print(f"FAIL: Component 2 — B{row} formula invalid: {'; '.join(reasons)} | value={val!r}")

            if valid_formula_count == len(formula_cells) and valid_formula_count == 11:
                print(f"PASS: Component 2 — All 11 formulas use MID+FIND referencing column A")
                total_score += 0.3
            elif valid_formula_count > 0:
                partial = round(0.3 * (valid_formula_count / 11), 3)
                print(f"PARTIAL: Component 2 — {valid_formula_count}/11 formulas use MID+FIND pattern (partial: {partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — No formulas use MID+FIND pattern referencing column A")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Each formula correctly references its own row number in column A (0.2 points)
    # e.g., B2 formula should reference A2, B3 should reference A3, etc.
    try:
        if not formula_cells:
            print("FAIL: Component 3 — No formula cells found (skipping row reference check)")
        else:
            correct_row_ref_count = 0
            for row, val in formula_cells:
                val_upper = val.upper().replace(' ', '')
                # Check that the formula references A{row} (same row)
                expected_ref = f'A{row}'
                if expected_ref in val_upper:
                    correct_row_ref_count += 1
                else:
                    print(f"FAIL: Component 3 — B{row} formula does not reference A{row}: {val!r}")

            if correct_row_ref_count == len(formula_cells) and correct_row_ref_count == 11:
                print(f"PASS: Component 3 — All 11 formulas correctly reference their own row in column A")
                total_score += 0.2
            elif correct_row_ref_count > 0:
                partial = round(0.2 * (correct_row_ref_count / 11), 3)
                print(f"PARTIAL: Component 3 — {correct_row_ref_count}/11 formulas have correct row references (partial: {partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No formulas have correct row references")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
