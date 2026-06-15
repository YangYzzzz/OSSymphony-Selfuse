"""
Initial Setup: Extract domain names from email addresses using MID/FIND formula
Task ID: calc_fma_mid_find_056
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'calc_fma_mid_find_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Emails ---
    ws = wb.active
    ws.title = 'Emails'

    # Headers
    ws['A1'] = 'Email'
    ws['B1'] = 'Domain'

    # Email data (rows 2-12, matching context exactly)
    emails = [
        'user1@gmail.com',
        'user2@yahoo.com',
        'user3@outlook.com',
        'user4@company.net',
        'user5@startup.io',
        'user6@enterprise.org',
        'user7@webmail.co.uk',
        'user8@proton.me',
        'user9@icloud.com',
        'user10@fastmail.com',
        'user11@zoho.com',
    ]

    for i, email in enumerate(emails, start=2):
        ws.cell(row=i, column=1, value=email)
        # Column B is intentionally empty (task requires filling it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
