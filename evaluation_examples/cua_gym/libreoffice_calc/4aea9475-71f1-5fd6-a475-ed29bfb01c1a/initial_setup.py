"""
Initial Setup: Create employee spreadsheet with plain numeric IDs
Task ID: calc_lf_082
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Employees ---
    ws = wb.active
    ws.title = 'Employees'

    # Headers
    headers = ['Name', 'Employee ID']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Employee data - realistic content with numeric IDs (NO custom format)
    employees = [
        ['Alice', 42],
        ['Bob', 1358],
        ['Carol', 100005],
    ]

    for r, (name, emp_id) in enumerate(employees, 2):
        name_cell = ws.cell(row=r, column=1, value=name)
        name_cell.font = Font(name='Calibri', size=11)
        name_cell.border = thin_border

        id_cell = ws.cell(row=r, column=2, value=emp_id)
        id_cell.font = Font(name='Calibri', size=11)
        id_cell.border = thin_border
        # Number format is 'General' by default - no custom format applied

    # Adjust column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
