"""
Reward Script: Apply custom number format "EMP-"000000 to employee ID cells
Task ID: calc_lf_082
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Each of B2:B4 has number_format '"EMP-"000000' (0.2 each, progressive)
  Component 2 (0.4): All three cells have the format AND values remain numeric integers (42, 1358, 100005)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_082'

EXPECTED_FORMAT = '"EMP-"000000'
EXPECTED_VALUES = {
    'B2': 42,
    'B3': 1358,
    'B4': 100005,
}


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Employees' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Employees']

    # Component 1: Each of B2:B4 has the custom number format (0.2 pts each, total 0.6)
    cells_with_format = []
    for coord in ['B2', 'B3', 'B4']:
        try:
            cell = ws[coord]
            nf = cell.number_format
            if nf == EXPECTED_FORMAT:
                print(f"PASS: Component 1 — {coord} has correct format '{nf}' (0.2 pts)")
                total_score += 0.2
                cells_with_format.append(coord)
            else:
                print(f"FAIL: Component 1 — {coord} format is '{nf}', expected '{EXPECTED_FORMAT}'")
        except Exception as e:
            print(f"ERROR: Component 1 — {coord}: {e}")

    # Component 2: All three cells have the format AND values remain as original numeric integers (0.4 pts)
    # This is a compound check: format must be applied AND values must be preserved as numeric.
    # On initial_env: format is 'General', so this fails (even though values are numeric).
    try:
        all_format_ok = len(cells_with_format) == 3
        all_values_ok = True

        for coord, expected_val in EXPECTED_VALUES.items():
            cell = ws[coord]
            val = cell.value
            if not isinstance(val, (int, float)):
                print(f"FAIL: Component 2 — {coord} value is not numeric: {val!r} (type={type(val).__name__})")
                all_values_ok = False
            elif int(val) != expected_val:
                print(f"FAIL: Component 2 — {coord} value is {val}, expected {expected_val}")
                all_values_ok = False

        if all_format_ok and all_values_ok:
            print(f"PASS: Component 2 — All cells have format applied AND values preserved as numeric (0.4 pts)")
            total_score += 0.4
        elif not all_format_ok:
            print(f"FAIL: Component 2 — Not all cells have the custom format (only {len(cells_with_format)}/3)")
        # If format is OK but values are wrong, the per-value FAIL messages above already printed
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist unsaved GUI state before verifying
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
