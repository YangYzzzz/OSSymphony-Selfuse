"""
Initial Setup: Calculate working days between start and end dates excluding weekends and public holidays
Task ID: calc_fma_networkdays_holiday_058
Domain: libreoffice_calc
"""

import os
import openpyxl
from datetime import date

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'calc_fma_networkdays_holiday_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: WorkDays ---
    ws = wb.active
    ws.title = 'WorkDays'

    # Row 1: Headers
    ws['A1'] = 'Start Date'
    ws['B1'] = 'End Date'
    ws['C1'] = 'Working Days'
    ws['D1'] = ''
    ws['E1'] = 'Public Holidays'

    # Column A: Start dates (rows 2-9)
    start_dates = [
        date(2024, 1, 2),
        date(2024, 2, 1),
        date(2024, 3, 15),
        date(2024, 4, 1),
        date(2024, 5, 20),
        date(2024, 6, 10),
        date(2024, 7, 1),
        date(2024, 8, 5),
    ]
    for i, d in enumerate(start_dates, start=2):
        ws.cell(row=i, column=1, value=d)
        ws.cell(row=i, column=1).number_format = 'yyyy-mm-dd'

    # Column B: End dates (rows 2-9)
    end_dates = [
        date(2024, 1, 31),
        date(2024, 2, 29),
        date(2024, 3, 29),
        date(2024, 4, 30),
        date(2024, 5, 31),
        date(2024, 6, 28),
        date(2024, 7, 31),
        date(2024, 8, 30),
    ]
    for i, d in enumerate(end_dates, start=2):
        ws.cell(row=i, column=2, value=d)
        ws.cell(row=i, column=2).number_format = 'yyyy-mm-dd'

    # Column C: Working Days header, cells C2:C9 EMPTY (task will fill these)
    # (already empty by default)

    # Column E: Public holiday dates (rows 2-8)
    holidays = [
        date(2024, 1, 1),
        date(2024, 1, 15),
        date(2024, 2, 19),
        date(2024, 5, 27),
        date(2024, 7, 4),
        date(2024, 9, 2),
        date(2024, 11, 28),
    ]
    for i, d in enumerate(holidays, start=2):
        ws.cell(row=i, column=5, value=d)
        ws.cell(row=i, column=5).number_format = 'yyyy-mm-dd'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
