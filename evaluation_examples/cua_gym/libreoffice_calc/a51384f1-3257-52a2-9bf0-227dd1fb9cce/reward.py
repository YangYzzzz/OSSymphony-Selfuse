"""
Reward Script: Calculate working days using NETWORKDAYS with holiday exclusion
Task ID: calc_fma_networkdays_holiday_058
Domain: libreoffice_calc
Scoring:
  Component 1: At least one NETWORKDAYS formula in C2:C9              — 0.3 pts
  Component 2: All 8 cells C2:C9 contain NETWORKDAYS formulas         — 0.4 pts
  Component 3: All formulas reference the holiday range $E$2:$E$8     — 0.3 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_networkdays_holiday_058'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'WorkDays' not in wb.sheetnames:
        print("FAIL: Sheet 'WorkDays' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['WorkDays']

    # Collect formula status for C2:C9
    formula_present = []   # True if NETWORKDAYS formula is present in this cell
    holiday_ref_present = []  # True if formula references $E$2:$E$8

    for row in range(2, 10):
        cell_val = ws.cell(row=row, column=3).value
        if cell_val is None:
            formula_present.append(False)
            holiday_ref_present.append(False)
            continue

        val_str = str(cell_val).strip().upper()

        # Check for NETWORKDAYS function usage
        has_networkdays = 'NETWORKDAYS' in val_str
        formula_present.append(has_networkdays)

        # Check for holiday range reference: $E$2:$E$8 (absolute or relative accepted,
        # but task context specifies $E$2:$E$8 as the correct form)
        # Accept both absolute and relative references to E2:E8
        has_holiday_ref = bool(
            re.search(r'\$?E\$?2:\$?E\$?8', val_str, re.IGNORECASE)
        )
        holiday_ref_present.append(has_holiday_ref and has_networkdays)

    count_with_formula = sum(formula_present)
    count_with_holiday = sum(holiday_ref_present)

    # Component 1: At least one NETWORKDAYS formula in C2:C9 (0.3 points)
    # This fails on the initial file (all cells None) and passes on the golden file.
    try:
        if count_with_formula >= 1:
            print(f"PASS: Component 1 — {count_with_formula}/8 cells have NETWORKDAYS formulas (0.3 pts)")
            total_score += 0.3
        else:
            print("FAIL: Component 1 — No NETWORKDAYS formulas found in C2:C9")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 8 cells C2:C9 contain NETWORKDAYS formulas (0.4 points)
    # Partial credit is already given in Component 1; this rewards full completion.
    try:
        if count_with_formula == 8:
            print("PASS: Component 2 — All 8 cells C2:C9 have NETWORKDAYS formulas (0.4 pts)")
            total_score += 0.4
        else:
            missing_rows = [i + 2 for i, ok in enumerate(formula_present) if not ok]
            print(f"FAIL: Component 2 — Only {count_with_formula}/8 cells have NETWORKDAYS formulas; missing rows: {missing_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All formulas reference the holiday range E2:E8 (0.3 points)
    # Ensures the holiday exclusion argument is correctly included.
    try:
        if count_with_holiday == 8:
            print("PASS: Component 3 — All 8 formulas reference the holiday range E2:E8 (0.3 pts)")
            total_score += 0.3
        else:
            missing_rows = [i + 2 for i, ok in enumerate(holiday_ref_present) if not ok]
            print(f"FAIL: Component 3 — Only {count_with_holiday}/8 formulas reference E2:E8; missing/incorrect rows: {missing_rows}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
