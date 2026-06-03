"""
Initial Setup: HR Absence Calendar — January 2025
Task ID: calc_hr_absence_calendar_054
Domain: libreoffice_calc

Creates 'January Absences' sheet with:
  - Row 1: headers (Employee + Jan 1 ... Jan 31)
  - Rows 2-35: 34 employees with absence codes (AL/SL/TR or empty)
  - Row 36: empty (for totals — to be filled in golden patch)
  - NO conditional formatting (task requirement)
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_absence_calendar_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'January Absences'

    # --- Row 1: Headers ---
    ws['A1'] = 'Employee'
    for day in range(1, 32):
        col = day + 1  # B=2 for Jan 1, AF=32 for Jan 31
        ws.cell(row=1, column=col, value=f'Jan {day}')

    # --- Employee list (realistic names, 34 employees) ---
    employees = [
        'Sarah Chen',
        'Marcus Johnson',
        'Priya Patel',
        'James Okafor',
        'Emma Lindström',
        'Daniel Rosenberg',
        'Fatima Al-Hassan',
        'Carlos Mendez',
        'Yuki Tanaka',
        'Amelia Burke',
        'Kwame Asante',
        'Isabelle Martin',
        'Ravi Sharma',
        'Natasha Ivanova',
        'David Nguyen',
        'Chloe Dubois',
        'Ahmed Khalil',
        'Olivia Campbell',
        'Liu Wei',
        'Benjamin Osei',
        'Mia Kowalski',
        'Tariq Rahman',
        'Sofia Hernandez',
        'Ethan Brooks',
        'Amara Diallo',
        'Lucas Schmidt',
        'Zara Abubakar',
        'Patrick Nkrumah',
        'Hannah Bergström',
        'Jamal Carter',
        'Nina Petrov',
        'Felix Wagner',
        'Aaliya Singh',
        'Thomas Dumont',
    ]

    # --- Absence data: realistic distribution of AL, SL, TR and empty cells ---
    import random
    random.seed(42)  # reproducible

    # Define absence codes and their approximate frequencies
    # ~65% present (empty), ~20% AL, ~10% SL, ~5% TR
    options = [''] * 65 + ['AL'] * 20 + ['SL'] * 10 + ['TR'] * 5

    for row_idx, name in enumerate(employees, start=2):
        ws.cell(row=row_idx, column=1, value=name)
        for col in range(2, 33):  # columns B through AF (31 days)
            code = random.choice(options)
            if code:
                ws.cell(row=row_idx, column=col, value=code)
            # empty cells are left as None (no value written)

    # Row 36 left empty (for Total Absences label and COUNTIF formulas)
    # A36 is empty, B36:AF36 are empty

    # Set column A width for employee names
    ws.column_dimensions['A'].width = 22

    # Set narrow width for day columns
    for col in range(2, 33):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col)].width = 6

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
