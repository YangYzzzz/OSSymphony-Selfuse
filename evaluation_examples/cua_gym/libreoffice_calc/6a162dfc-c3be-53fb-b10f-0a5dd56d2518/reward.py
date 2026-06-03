"""
Reward Script: Monthly Absence Calendar — Color-coding and Totals Row
Task ID: calc_hr_absence_calendar_054
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Conditional formatting — AL rule (blue #4472C4) on B2:AF35    — 0.15 pts
  Component 2: Conditional formatting — SL rule (yellow #FFFF00) on B2:AF35  — 0.15 pts
  Component 3: Conditional formatting — TR rule (green #70AD47) on B2:AF35   — 0.15 pts
  Component 4: A36 = 'Total Absences' (bold)                                  — 0.15 pts
  Component 5: B36:AF36 contain COUNTIF formulas (spot-check + full check)   — 0.40 pts
  Total: 1.0
"""

import os
import re
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_absence_calendar_054'
SHEET_NAME = 'January Absences'


def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, remove whitespace."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def check_countif_formula(formula, col_letter):
    """
    Check if formula matches the expected COUNTIF pattern for a given column.
    Expected: =COUNTIF(<COL>2:<COL>35,"AL")+COUNTIF(<COL>2:<COL>35,"SL")+COUNTIF(<COL>2:<COL>35,"TR")
    Accepts case-insensitive and whitespace-normalized forms.
    """
    if not isinstance(formula, str):
        return False
    norm = normalize_formula(formula)
    col = col_letter.upper()
    # Build expected pattern (normalized)
    expected = (
        f'=COUNTIF({col}2:{col}35,"AL")'
        f'+COUNTIF({col}2:{col}35,"SL")'
        f'+COUNTIF({col}2:{col}35,"TR")'
    ).upper().replace(' ', '')
    return norm == expected


def verify_task(file_path):
    """
    Verify task completion for the HR Absence Calendar task.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists — precondition gate
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------------------------
    # Collect conditional formatting rules on the target range B2:AF35
    # -----------------------------------------------------------------------
    cf_rules_on_range = []
    for cf in ws.conditional_formatting:
        cf_str = str(cf)
        # Match the range B2:AF35 (may appear with or without $ or extra spaces)
        if 'B2:AF35' in cf_str.upper() or 'B2:AF35' in cf_str:
            for rule in cf.rules:
                cf_rules_on_range.append(rule)

    # Build lookup: formula expression -> fill color
    cf_fill_map = {}
    cf_font_color_map = {}
    for rule in cf_rules_on_range:
        if rule.type == 'expression' and hasattr(rule, 'formula') and rule.formula:
            expr = normalize_formula(rule.formula[0]) if rule.formula else ''
            if hasattr(rule, 'dxf') and rule.dxf:
                dxf = rule.dxf
                if dxf.fill:
                    try:
                        rgb = dxf.fill.fgColor.rgb
                        cf_fill_map[expr] = rgb
                    except Exception:
                        pass
                if dxf.font:
                    try:
                        font_rgb = dxf.font.color.rgb
                        cf_font_color_map[expr] = font_rgb
                    except Exception:
                        pass

    # Expected CF expressions (normalized)
    # Formula references the top-left cell of the range
    al_expr = normalize_formula('B2="AL"')
    sl_expr = normalize_formula('B2="SL"')
    tr_expr = normalize_formula('B2="TR"')

    # Component 1: AL rule — background blue #4472C4 (ARGB: FF4472C4) (0.15 pts)
    try:
        al_fill = cf_fill_map.get(al_expr, '')
        al_font = cf_font_color_map.get(al_expr, '')
        # Accept FF4472C4 for fill; font color should be white (FFFFFF or 00FFFFFF)
        al_fill_ok = al_fill.upper().endswith('4472C4') if al_fill else False
        # Font white: either 00FFFFFF or FFFFFFFF are acceptable
        al_font_ok = ('FFFFFF' in al_font.upper()) if al_font else False

        if al_fill_ok:
            print(f"PASS: Component 1a — AL fill color correct: {al_fill} (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 1a — AL fill expected *4472C4, found: {repr(al_fill)}")

        if al_font_ok:
            print(f"PASS: Component 1b — AL font color is white: {al_font} (0.03 pts)")
            total_score += 0.03
        else:
            print(f"FAIL: Component 1b — AL font color expected white (*FFFFFF), found: {repr(al_font)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SL rule — background yellow #FFFF00 (ARGB: FFFFFF00) (0.15 pts)
    try:
        sl_fill = cf_fill_map.get(sl_expr, '')
        sl_fill_ok = sl_fill.upper().endswith('FFFF00') if sl_fill else False
        if sl_fill_ok:
            print(f"PASS: Component 2 — SL fill color correct: {sl_fill} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — SL fill expected *FFFF00, found: {repr(sl_fill)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: TR rule — background green #70AD47 (ARGB: FF70AD47) (0.15 pts)
    try:
        tr_fill = cf_fill_map.get(tr_expr, '')
        tr_fill_ok = tr_fill.upper().endswith('70AD47') if tr_fill else False
        if tr_fill_ok:
            print(f"PASS: Component 3 — TR fill color correct: {tr_fill} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — TR fill expected *70AD47, found: {repr(tr_fill)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: A36 = 'Total Absences' (bold) (0.15 pts)
    try:
        a36_val = ws.cell(row=36, column=1).value
        a36_bold = ws.cell(row=36, column=1).font.bold

        a36_text_ok = (isinstance(a36_val, str) and a36_val.strip().lower() == 'total absences')
        a36_bold_ok = (a36_bold is True)

        if a36_text_ok and a36_bold_ok:
            print(f"PASS: Component 4 — A36='{a36_val}' bold={a36_bold} (0.15 pts)")
            total_score += 0.15
        elif a36_text_ok and not a36_bold_ok:
            # Partial: label present but not bold — give 0.10
            print(f"PARTIAL: Component 4 — A36='{a36_val}' but bold={a36_bold}. Giving 0.10 pts.")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — A36 expected 'Total Absences' (bold), found: {repr(a36_val)}, bold={a36_bold}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: B36:AF36 contain COUNTIF formulas (0.40 pts)
    # Spot-check first column (B36), last column (AF36), and two middle columns (L36, U36)
    # Full pass requires all 31 columns to have correct COUNTIF formulas
    try:
        spot_cols = [
            (2, 'B'),
            (12, 'L'),
            (21, 'U'),
            (32, 'AF'),
        ]
        spot_pass = 0
        for col_idx, col_letter in spot_cols:
            val = ws.cell(row=36, column=col_idx).value
            if check_countif_formula(val, col_letter):
                spot_pass += 1
            else:
                print(f"  FAIL spot-check: {col_letter}36 = {repr(val)}")

        # Full check: all 31 columns B-AF (cols 2-32)
        full_pass = 0
        full_fail_list = []
        for col_idx in range(2, 33):
            col_letter = get_column_letter(col_idx)
            val = ws.cell(row=36, column=col_idx).value
            if check_countif_formula(val, col_letter):
                full_pass += 1
            else:
                full_fail_list.append(f"{col_letter}36={repr(val)}")

        if full_pass == 31:
            print(f"PASS: Component 5 — All 31 COUNTIF formulas correct in B36:AF36 (0.40 pts)")
            total_score += 0.40
        elif full_pass >= 26:
            # Mostly correct (>=26/31): award proportional credit
            if full_fail_list:
                print(f"  Failed cells: {', '.join(full_fail_list[:5])}" + (" ..." if len(full_fail_list) > 5 else ""))
            score_5 = round(0.40 * full_pass / 31, 2)
            print(f"PARTIAL: Component 5 — {full_pass}/31 COUNTIF formulas correct ({score_5} pts)")
            if full_pass >= 26:
                total_score += score_5
        elif full_pass > 0:
            # Fewer than 26 correct: award proportional credit if any pass
            score_5 = round(0.40 * full_pass / 31, 2)
            print(f"PARTIAL: Component 5 — {full_pass}/31 formulas correct ({score_5} pts). Spot: {spot_pass}/4")
            if full_pass > 0:
                total_score += score_5
        else:
            print(f"FAIL: Component 5 — Only {full_pass}/31 COUNTIF formulas correct. Spot: {spot_pass}/4")
            if full_fail_list:
                print(f"  Sample failures: {', '.join(full_fail_list[:5])}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
