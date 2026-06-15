"""
Initial Setup: Student enrollment list with empty Student Number column
Task ID: osworld_calc_fill_sequence_numbers_003
Domain: libreoffice_calc

Creates a student enrollment spreadsheet where column A (Student Number)
is labeled but empty — the agent must fill it with STU-0001 to STU-0035.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Student Enrollment ---
    ws = wb.active
    ws.title = "Enrollment"

    # Headers
    headers = ['Student Number', 'Full Name', 'Major', 'Year', 'GPA']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Student data — 35 realistic students (rows 2-36)
    # Column A is intentionally left EMPTY (no student numbers)
    students = [
        ('Emily Zhang',       'Computer Science',    2, 3.85),
        ('Marcus Thompson',   'Biology',             3, 3.62),
        ('Sophia Reyes',      'Business Admin',      1, 3.74),
        ('Liam O\'Brien',     'Mechanical Eng',      4, 3.41),
        ('Ava Patel',         'Psychology',          2, 3.93),
        ('Noah Williams',     'History',             3, 3.55),
        ('Isabella Kim',      'Nursing',             4, 3.78),
        ('James Carter',      'Computer Science',    1, 3.60),
        ('Mia Gonzalez',      'English Literature',  2, 3.82),
        ('Ethan Brown',       'Civil Engineering',   3, 3.47),
        ('Charlotte Davis',   'Chemistry',           4, 3.71),
        ('Alexander Lee',     'Economics',           1, 3.89),
        ('Amelia Wilson',     'Political Science',   2, 3.53),
        ('Benjamin Harris',   'Mathematics',         3, 3.95),
        ('Harper Martinez',   'Art History',         4, 3.38),
        ('Daniel Anderson',   'Physics',             1, 3.66),
        ('Ella Thomas',       'Sociology',           2, 3.77),
        ('Matthew Jackson',   'Accounting',          3, 3.44),
        ('Scarlett White',    'Biomedical Eng',      4, 3.88),
        ('Henry Taylor',      'Philosophy',          1, 3.50),
        ('Luna Moore',        'Communications',      2, 3.72),
        ('Jackson Martin',    'Finance',             3, 3.61),
        ('Aria Garcia',       'Environmental Sci',   4, 3.83),
        ('Sebastian Lewis',   'Computer Science',    1, 3.69),
        ('Penelope Walker',   'Nutrition',           2, 3.56),
        ('Aiden Hall',        'Statistics',          3, 3.91),
        ('Chloe Young',       'Public Health',       4, 3.43),
        ('Owen Allen',        'Marketing',           1, 3.75),
        ('Layla King',        'Data Science',        2, 3.87),
        ('Ryan Hernandez',    'Architecture',        3, 3.58),
        ('Zoey Wright',       'Music Theory',        4, 3.40),
        ('Nathan Scott',      'Software Eng',        1, 3.80),
        ('Lily Green',        'Linguistics',         2, 3.64),
        ('Julian Baker',      'Information Sys',     3, 3.92),
        ('Stella Adams',      'Graphic Design',      4, 3.67),
    ]

    for row_idx, (name, major, year, gpa) in enumerate(students, 2):
        # Column A (Student Number): intentionally EMPTY
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=major)
        ws.cell(row=row_idx, column=4, value=year)
        ws.cell(row=row_idx, column=5, value=gpa)

    # Column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
