"""
Reward Script: Fill student number sequence in column A
Task ID: osworld_calc_fill_sequence_numbers_003
Domain: libreoffice_calc
Scoring:
  Component 1: First student number A2 == 'STU-0001'                   — 0.4 pts
  Component 2: Last student number A36 == 'STU-0035'                   — 0.3 pts
  Component 3: All 35 values A2:A36 match expected STU-XXXX pattern    — 0.3 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_003'


def verify_task(file_path):
    """
    Verify that column A of the Enrollment sheet contains sequential student
    numbers 'STU-0001' through 'STU-0035' in rows 2-36.

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Enrollment sheet must exist
    if 'Enrollment' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Enrollment' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Enrollment']

    # Component 1: First student number A2 == 'STU-0001' (0.4 points)
    # This FAILS on initial_env (A2 is None) and PASSES on golden_env
    try:
        a2_value = ws.cell(row=2, column=1).value
        if a2_value is not None and str(a2_value).strip() == 'STU-0001':
            print(f"PASS: Component 1 -- A2 == 'STU-0001' (value: {a2_value}) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- Expected A2='STU-0001', found: {a2_value!r}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Last student number A36 == 'STU-0035' (0.3 points)
    # This FAILS on initial_env (A36 is None) and PASSES on golden_env
    try:
        a36_value = ws.cell(row=36, column=1).value
        if a36_value is not None and str(a36_value).strip() == 'STU-0035':
            print(f"PASS: Component 2 -- A36 == 'STU-0035' (value: {a36_value}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 -- Expected A36='STU-0035', found: {a36_value!r}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: All 35 values A2:A36 match expected 'STU-XXXX' pattern (0.3 points)
    # Expected values: STU-0001 to STU-0035 in order
    # This FAILS on initial_env (all None) and PASSES on golden_env
    try:
        expected_values = [f"STU-{i:04d}" for i in range(1, 36)]
        mismatch_count = 0
        for idx, expected in enumerate(expected_values):
            actual = ws.cell(row=idx + 2, column=1).value
            actual_str = str(actual).strip() if actual is not None else None
            if actual_str != expected:
                mismatch_count += 1

        if mismatch_count == 0:
            print(f"PASS: Component 3 -- All 35 student numbers A2:A36 match 'STU-0001'-'STU-0035' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 -- {mismatch_count}/35 values do not match expected pattern")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
