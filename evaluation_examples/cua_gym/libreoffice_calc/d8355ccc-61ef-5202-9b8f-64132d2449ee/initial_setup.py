"""
Initial Setup: Create a spreadsheet with customer complaints data by category
Task ID: calc_chart_pie_3d_066
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_pie_3d_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Complaints ---
    ws = wb.active
    ws.title = 'Complaints'

    # Headers
    ws['A1'] = 'Category'
    ws['B1'] = 'Count'

    # Style headers bold
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # Data rows - realistic complaint categories and counts
    data = [
        ('Product Defect', 124),
        ('Late Delivery', 89),
        ('Wrong Item', 67),
        ('Billing Issue', 45),
        ('Customer Service', 38),
        ('Other', 28),
    ]

    for r, (category, count) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=category)
        ws.cell(row=r, column=2, value=count)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10

    # NO charts in initial file - agent must create the chart
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Complaints')
    print('Rows: 1 header + 6 data rows')
    print('No charts present (agent must create the 3D pie chart)')


create_initial()
