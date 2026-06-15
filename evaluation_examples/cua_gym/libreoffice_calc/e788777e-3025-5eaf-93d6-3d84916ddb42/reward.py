"""
Reward Script: Compare procurement forecast vs actual purchase orders by month/category.
Task ID: calc_ops_supply_chain_forecast_vs_actual_066
Domain: libreoffice_calc

Scoring Rubric:
- Component 1: Variance $ formulas in E2:E61 (=D-C pattern)      0.25 pts
- Component 2: Variance % formulas in F2:F61 (=E/C pattern)      0.20 pts
- Component 3: Accuracy status formulas in G2:G61 (IF with thresholds) 0.25 pts
- Component 4: Named ranges OverOrderThreshold + UnderOrderThreshold   0.15 pts
- Component 5: Conditional formatting on G2:G61 (3 color rules)   0.10 pts
- Component 6: Summary COUNTIF section (rows 63-66)                0.05 pts
Total: 1.00
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_forecast_vs_actual_066'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'ForecastVsActual' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ForecastVsActual' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ForecastVsActual']

    # ------------------------------------------------------------------
    # Component 1: Variance $ formulas in E2:E61 (=D{n}-C{n} pattern)
    # Expected: each cell En = =Dn-Cn  (0.25 points)
    # ------------------------------------------------------------------
    try:
        e_formula_count = 0
        e_total = 0
        for row in range(2, 62):  # rows 2..61 inclusive = 60 rows
            e_total += 1
            val = ws.cell(row=row, column=5).value  # column E
            if val is not None and isinstance(val, str):
                # Accept =Dn-Cn or =D{n}-C{n}
                formula = val.strip().upper().replace(' ', '')
                expected = f'=D{row}-C{row}'
                if formula == expected:
                    e_formula_count += 1
        if e_formula_count == 60:
            print(f"PASS: Component 1 — All 60 Variance $ formulas present in E2:E61 (0.25 pts)")
            total_score += 0.25
        elif e_formula_count >= 30:
            partial = 0.12
            print(f"PARTIAL: Component 1 — {e_formula_count}/60 Variance $ formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {e_formula_count}/60 Variance $ formulas found in E2:E61")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Variance % formulas in F2:F61 (=E{n}/C{n} pattern)
    # Expected: each cell Fn = =En/Cn  (0.20 points)
    # ------------------------------------------------------------------
    try:
        f_formula_count = 0
        for row in range(2, 62):
            val = ws.cell(row=row, column=6).value  # column F
            if val is not None and isinstance(val, str):
                formula = val.strip().upper().replace(' ', '')
                expected = f'=E{row}/C{row}'
                if formula == expected:
                    f_formula_count += 1
        if f_formula_count == 60:
            print(f"PASS: Component 2 — All 60 Variance % formulas present in F2:F61 (0.20 pts)")
            total_score += 0.20
        elif f_formula_count >= 30:
            partial = 0.10
            print(f"PARTIAL: Component 2 — {f_formula_count}/60 Variance % formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {f_formula_count}/60 Variance % formulas found in F2:F61")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Accuracy status formulas in G2:G61
    # Expected: =IF(Fn>0.2,"OVER-ORDERED",IF(Fn<-0.15,"UNDER-ORDERED","ON TARGET"))
    # (0.25 points)
    # ------------------------------------------------------------------
    try:
        g_formula_count = 0
        for row in range(2, 62):
            val = ws.cell(row=row, column=7).value  # column G
            if val is not None and isinstance(val, str):
                formula = val.strip().upper().replace(' ', '')
                # Accept various equivalent forms
                # Core pattern: IF with F>0.2 OVER-ORDERED, F<-0.15 UNDER-ORDERED, ON TARGET
                has_over = 'OVER-ORDERED' in formula
                has_under = 'UNDER-ORDERED' in formula
                has_on_target = 'ONTARGET' in formula or 'ON TARGET' in val.upper()
                has_f_ref = f'F{row}' in formula
                has_threshold_over = '0.2' in formula or '20%' in formula
                has_threshold_under = '-0.15' in formula or '-15%' in formula
                if (has_over and has_under and has_on_target and has_f_ref and
                        has_threshold_over and has_threshold_under):
                    g_formula_count += 1
        if g_formula_count == 60:
            print(f"PASS: Component 3 — All 60 Accuracy IF formulas present in G2:G61 (0.25 pts)")
            total_score += 0.25
        elif g_formula_count >= 30:
            partial = 0.12
            print(f"PARTIAL: Component 3 — {g_formula_count}/60 Accuracy formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {g_formula_count}/60 Accuracy formulas found in G2:G61")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: Named ranges OverOrderThreshold and UnderOrderThreshold
    # (0.15 points — 0.075 each)
    # ------------------------------------------------------------------
    try:
        defined_names = dict(wb.defined_names.items())
        has_over_threshold = 'OverOrderThreshold' in defined_names
        has_under_threshold = 'UnderOrderThreshold' in defined_names

        named_score = 0.0
        if has_over_threshold:
            named_score += 0.075
            print(f"PASS: Component 4a — Named range 'OverOrderThreshold' exists")
        else:
            print(f"FAIL: Component 4a — Named range 'OverOrderThreshold' not found")

        if has_under_threshold:
            named_score += 0.075
            print(f"PASS: Component 4b — Named range 'UnderOrderThreshold' exists")
        else:
            print(f"FAIL: Component 4b — Named range 'UnderOrderThreshold' not found")

        if named_score > 0:
            print(f"  Component 4 total: {named_score:.3f} pts")
            total_score += named_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ------------------------------------------------------------------
    # Component 5: Conditional formatting on G2:G61 — 3 color rules
    # Expected: OVER-ORDERED (red/pink), UNDER-ORDERED (orange/yellow), ON TARGET (green)
    # (0.10 points)
    # ------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        g_ranges = []
        for cf in cf_rules:
            cf_str = str(cf)
            if 'G' in cf_str.upper():
                g_ranges.append(cf)

        if not g_ranges:
            print(f"FAIL: Component 5 — No conditional formatting found on column G")
        else:
            # Count rules that reference OVER-ORDERED, UNDER-ORDERED, ON TARGET
            over_cf_count = 0
            under_cf_count = 0
            on_target_cf_count = 0
            for cf in g_ranges:
                for rule in cf.rules:
                    formula_str = str(rule.formula).upper() if rule.formula else ''
                    if 'OVER-ORDERED' in formula_str:
                        over_cf_count += 1
                    if 'UNDER-ORDERED' in formula_str:
                        under_cf_count += 1
                    if 'ON TARGET' in formula_str:
                        on_target_cf_count += 1

            cf_count = sum([over_cf_count > 0, under_cf_count > 0, on_target_cf_count > 0])
            if cf_count == 3:
                print(f"PASS: Component 5 — All 3 conditional formatting rules present on G2:G61 (0.10 pts)")
                total_score += 0.10
            elif cf_count >= 1:
                partial = round(cf_count * 0.033, 3)
                print(f"PARTIAL: Component 5 — {cf_count}/3 CF rules present ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 — No matching CF rules found referencing status labels")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ------------------------------------------------------------------
    # Component 6: Summary COUNTIF section (rows 63-66)
    # Expected:
    #   Row 63: 'Summary' label in column A
    #   Row 64: 'OVER-ORDERED count' in A, =COUNTIF(G2:G61,"OVER-ORDERED") in B
    #   Row 65: 'UNDER-ORDERED count' in A, =COUNTIF(G2:G61,"UNDER-ORDERED") in B
    #   Row 66: 'ON TARGET count' in A, =COUNTIF(G2:G61,"ON TARGET") in B
    # (0.05 points)
    # ------------------------------------------------------------------
    try:
        summary_label = ws.cell(row=63, column=1).value
        over_label = ws.cell(row=64, column=1).value
        over_formula = ws.cell(row=64, column=2).value
        under_label = ws.cell(row=65, column=1).value
        under_formula = ws.cell(row=65, column=2).value
        on_target_label = ws.cell(row=66, column=1).value
        on_target_formula = ws.cell(row=66, column=2).value

        has_summary = summary_label is not None and 'SUMMARY' in str(summary_label).upper()
        has_over_countif = (over_formula is not None and isinstance(over_formula, str) and
                            'COUNTIF' in over_formula.upper() and 'OVER-ORDERED' in over_formula.upper())
        has_under_countif = (under_formula is not None and isinstance(under_formula, str) and
                             'COUNTIF' in under_formula.upper() and 'UNDER-ORDERED' in under_formula.upper())
        has_on_target_countif = (on_target_formula is not None and isinstance(on_target_formula, str) and
                                  'COUNTIF' in on_target_formula.upper() and 'ON TARGET' in on_target_formula.upper())

        checks_passed = sum([has_summary, has_over_countif, has_under_countif, has_on_target_countif])
        if checks_passed >= 3:
            print(f"PASS: Component 6 — Summary COUNTIF section present in rows 63-66 (0.05 pts)")
            total_score += 0.05
        elif checks_passed >= 1:
            partial = 0.02
            print(f"PARTIAL: Component 6 — {checks_passed}/4 summary elements present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Summary COUNTIF section not found in rows 63-66")
            print(f"  Row63 A={repr(summary_label)}, Row64 B={repr(over_formula)}, Row65 B={repr(under_formula)}, Row66 B={repr(on_target_formula)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
