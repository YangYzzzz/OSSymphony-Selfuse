"""
Initial Setup: Supply Chain Forecast vs Actual
Task ID: calc_ops_supply_chain_forecast_vs_actual_066
Domain: libreoffice_calc

Creates a spreadsheet with 60 month-category procurement rows.
Columns A-D are filled; E (Variance $), F (Variance %), G (Accuracy) are empty.
No formulas, no conditional formatting, no named ranges.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_forecast_vs_actual_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ForecastVsActual'

    # --- Headers (Row 1) ---
    headers = [
        'Month', 'Category',
        'Forecasted PO Value', 'Actual PO Value',
        'Variance $', 'Variance %', 'Accuracy'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Define 5 procurement categories
    categories = [
        'Raw Materials',
        'Packaging',
        'Electronics',
        'Logistics',
        'Maintenance & Repair'
    ]

    # Define 12 months (2024)
    months = [
        'Jan 2024', 'Feb 2024', 'Mar 2024', 'Apr 2024',
        'May 2024', 'Jun 2024', 'Jul 2024', 'Aug 2024',
        'Sep 2024', 'Oct 2024', 'Nov 2024', 'Dec 2024'
    ]

    # Seed base forecasted values per category (USD)
    base_forecast = {
        'Raw Materials':        120000,
        'Packaging':             45000,
        'Electronics':          210000,
        'Logistics':             78000,
        'Maintenance & Repair':  32000,
    }

    # Monthly variation multipliers (seasonal patterns)
    monthly_factor = {
        'Jan 2024': 0.88, 'Feb 2024': 0.92, 'Mar 2024': 1.05,
        'Apr 2024': 1.02, 'May 2024': 1.08, 'Jun 2024': 1.12,
        'Jul 2024': 0.95, 'Aug 2024': 0.98, 'Sep 2024': 1.10,
        'Oct 2024': 1.15, 'Nov 2024': 1.20, 'Dec 2024': 1.18,
    }

    # Actual variation vs forecast (some months over, some under)
    actual_deltas = {
        # (month_index, category_index): multiplier applied to forecast
        (0,  'Raw Materials'):        0.94,
        (0,  'Packaging'):            1.25,
        (0,  'Electronics'):          0.98,
        (0,  'Logistics'):            0.88,
        (0,  'Maintenance & Repair'): 1.05,
        (1,  'Raw Materials'):        1.00,
        (1,  'Packaging'):            0.82,
        (1,  'Electronics'):          1.22,
        (1,  'Logistics'):            1.03,
        (1,  'Maintenance & Repair'): 0.90,
        (2,  'Raw Materials'):        1.18,
        (2,  'Packaging'):            1.00,
        (2,  'Electronics'):          0.95,
        (2,  'Logistics'):            1.25,
        (2,  'Maintenance & Repair'): 0.87,
        (3,  'Raw Materials'):        0.97,
        (3,  'Packaging'):            1.05,
        (3,  'Electronics'):          1.10,
        (3,  'Logistics'):            0.93,
        (3,  'Maintenance & Repair'): 1.00,
        (4,  'Raw Materials'):        1.05,
        (4,  'Packaging'):            0.78,
        (4,  'Electronics'):          1.28,
        (4,  'Logistics'):            1.15,
        (4,  'Maintenance & Repair'): 0.96,
        (5,  'Raw Materials'):        1.22,
        (5,  'Packaging'):            1.12,
        (5,  'Electronics'):          0.91,
        (5,  'Logistics'):            1.05,
        (5,  'Maintenance & Repair'): 1.18,
        (6,  'Raw Materials'):        0.85,
        (6,  'Packaging'):            0.95,
        (6,  'Electronics'):          1.02,
        (6,  'Logistics'):            0.80,
        (6,  'Maintenance & Repair'): 1.00,
        (7,  'Raw Materials'):        1.00,
        (7,  'Packaging'):            1.08,
        (7,  'Electronics'):          0.93,
        (7,  'Logistics'):            1.20,
        (7,  'Maintenance & Repair'): 0.88,
        (8,  'Raw Materials'):        1.15,
        (8,  'Packaging'):            0.90,
        (8,  'Electronics'):          1.05,
        (8,  'Logistics'):            0.98,
        (8,  'Maintenance & Repair'): 1.24,
        (9,  'Raw Materials'):        1.28,
        (9,  'Packaging'):            1.18,
        (9,  'Electronics'):          0.88,
        (9,  'Logistics'):            1.10,
        (9,  'Maintenance & Repair'): 0.92,
        (10, 'Raw Materials'):        1.05,
        (10, 'Packaging'):            0.85,
        (10, 'Electronics'):          1.30,
        (10, 'Logistics'):            1.02,
        (10, 'Maintenance & Repair'): 1.08,
        (11, 'Raw Materials'):        0.92,
        (11, 'Packaging'):            1.10,
        (11, 'Electronics'):          1.12,
        (11, 'Logistics'):            0.88,
        (11, 'Maintenance & Repair'): 1.22,
    }

    row = 2
    for mi, month in enumerate(months):
        for cat in categories:
            forecast_val = round(base_forecast[cat] * monthly_factor[month], 2)
            actual_mult  = actual_deltas.get((mi, cat), 1.0)
            actual_val   = round(forecast_val * actual_mult, 2)

            ws.cell(row=row, column=1, value=month)
            ws.cell(row=row, column=2, value=cat)
            ws.cell(row=row, column=3, value=forecast_val)
            ws.cell(row=row, column=4, value=actual_val)
            # Columns E, F, G intentionally left empty
            row += 1

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Rows of data: {row - 2}')
    print(f'  Columns E/F/G are empty (no formulas)')

create_initial()
