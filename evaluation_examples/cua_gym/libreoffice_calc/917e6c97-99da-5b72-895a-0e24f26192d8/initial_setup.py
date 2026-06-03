"""
Initial Setup: Fix #VALUE! error in addition formula
Task ID: calc_tbl_006
Domain: libreoffice_calc

Creates a spreadsheet with employee project cost data.
Row 3 has B3='N/A' (text) which causes F3's addition formula to show #VALUE!.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Costs"

    # --- Headers ---
    headers = ['Materials', 'Labor', 'Equipment', 'Project', 'Manager', 'Total Cost', 'Notes']
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data rows ---
    # Columns: A=Materials, B=Labor, C=Equipment, D=Project, E=Manager, F=Total (formula), G=Notes
    data = [
        # row 2
        [15000, 8500,  3200, 'Office Renovation',     'Sarah Chen',      None, ''],
        # row 3 — B3 will be 'N/A' text, causing #VALUE! in F3
        [100,   'N/A', 200,  'Server Migration',      'James Rivera',    None, ''],
        # row 4
        [22000, 14000, 6800, 'Warehouse Expansion',   'Priya Patel',     None, ''],
        # row 5
        [4500,  3200,  1100, 'Network Upgrade',       'Marcus Johnson',  None, ''],
        # row 6
        [31000, 19500, 8700, 'Data Center Build',     'Elena Kowalski',  None, ''],
        # row 7
        [7800,  5600,  2400, 'Security System',       'David Okonkwo',   None, ''],
        # row 8
        [12500, 9100,  4300, 'Parking Lot Repaving',  'Lisa Tanaka',     None, ''],
        # row 9
        [18000, 11000, 5500, 'HVAC Replacement',      'Robert Williams', None, ''],
        # row 10
        [9200,  6800,  2900, 'Lobby Redesign',        'Anika Sharma',    None, ''],
        # row 11
        [25000, 16000, 7200, 'Lab Equipment Install',  'Carlos Mendez',  None, ''],
        # row 12
        [6300,  4100,  1800, 'Fire Alarm Upgrade',    'Fatima Al-Hassan', None, ''],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            if val is not None and val != '':
                ws.cell(row=r, column=c, value=val)

    # --- Total Cost formulas in column F ---
    for r in range(2, 13):
        ws.cell(row=r, column=6, value=f'=A{r}+B{r}+C{r}')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 28

    # --- Number format for cost columns ---
    for r in range(2, 13):
        for c in [1, 2, 3, 6]:
            cell = ws.cell(row=r, column=c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
