"""
Reward Script: Replace B3 'N/A' with 0 and add note in G3
Task ID: calc_tbl_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): B3 is numeric 0 (was text 'N/A')
  Component 2 (0.3): F3 formula preserved (=A3+B3+C3)
  Component 3 (0.3): G3 contains 'Data unavailable for B'
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_006'


def persist_app_state(domain: str):
    """Best-effort save via Ctrl+S for any unsaved GUI edits."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: B3 is numeric 0 (0.4 points)
    # Initial state: B3 = 'N/A' (text). Golden state: B3 = 0 (number).
    try:
        b3_val = ws['B3'].value
        if isinstance(b3_val, (int, float)) and b3_val == 0:
            print(f"PASS: Component 1 — B3 is numeric 0 (found: {repr(b3_val)}) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — B3 should be numeric 0, found: {repr(b3_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F3 formula preserved as =A3+B3+C3 (0.3 points)
    # The formula should remain intact after replacing B3.
    # This component scores the fact that B3 change makes F3 calculate correctly.
    # On initial_env, F3 has the formula but B3='N/A' causes #VALUE! error.
    # On golden_env, F3 has the formula and B3=0 makes it compute to 300.
    # We check that the formula is still there AND B3 is numeric (compound check).
    try:
        f3_val = ws['F3'].value
        b3_val = ws['B3'].value
        if (isinstance(f3_val, str) and
                f3_val.upper().replace(" ", "") == "=A3+B3+C3" and
                isinstance(b3_val, (int, float))):
            print(f"PASS: Component 2 — F3 formula intact and B3 numeric, formula can compute (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — F3={repr(f3_val)}, B3={repr(b3_val)}; need formula + numeric B3")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G3 contains 'Data unavailable for B' (0.3 points)
    # Initial state: G3 is empty. Golden state: G3 = 'Data unavailable for B'.
    try:
        g3_val = ws['G3'].value
        if g3_val is not None and str(g3_val).strip() == 'Data unavailable for B':
            print(f"PASS: Component 3 — G3 note correct: {repr(g3_val)} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — G3 should be 'Data unavailable for B', found: {repr(g3_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
