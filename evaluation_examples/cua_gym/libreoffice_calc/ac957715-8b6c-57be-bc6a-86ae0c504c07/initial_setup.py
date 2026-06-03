"""
Initial Setup: Supply Chain Demand Planning — Create initial file with historical demand data
Task ID: calc_ops_supply_chain_demand_planning_056
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_demand_planning_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: DemandHistory                                               #
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = 'DemandHistory'

    # Header row
    headers_dh = ['SKU', 'Month-6', 'Month-5', 'Month-4', 'Month-3', 'Month-2', 'Month-1', 'Forecast Next Month']
    for col, h in enumerate(headers_dh, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # 30 SKU records with realistic historical demand data
    # Columns: SKU, Month-6, Month-5, Month-4, Month-3, Month-2, Month-1
    # Column H (Forecast Next Month) is intentionally left EMPTY
    sku_data = [
        ['SKU-A001', 1420, 1385, 1450, 1510, 1490, 1530],
        ['SKU-A002',  870,  910,  880,  920,  895,  940],
        ['SKU-A003', 2340, 2280, 2400, 2350, 2410, 2375],
        ['SKU-B001',  560,  590,  575,  610,  625,  600],
        ['SKU-B002', 1180, 1210, 1195, 1230, 1260, 1245],
        ['SKU-B003', 3200, 3150, 3280, 3310, 3250, 3340],
        ['SKU-C001',  780,  810,  795,  830,  820,  845],
        ['SKU-C002', 2100, 2050, 2180, 2160, 2200, 2175],
        ['SKU-C003',  460,  490,  475,  500,  515,  505],
        ['SKU-D001', 1650, 1620, 1700, 1680, 1720, 1695],
        ['SKU-D002',  325,  350,  340,  360,  370,  355],
        ['SKU-D003', 4100, 4050, 4200, 4180, 4250, 4230],
        ['SKU-E001', 1050, 1080, 1065, 1100, 1090, 1115],
        ['SKU-E002',  690,  720,  705,  740,  730,  755],
        ['SKU-E003', 2800, 2750, 2900, 2870, 2920, 2895],
        ['SKU-F001',  940,  970,  955,  990, 1005,  985],
        ['SKU-F002', 1780, 1750, 1820, 1800, 1850, 1830],
        ['SKU-F003',  230,  250,  240,  260,  270,  255],
        ['SKU-G001', 3500, 3450, 3600, 3570, 3620, 3590],
        ['SKU-G002', 1320, 1290, 1360, 1340, 1380, 1365],
        ['SKU-G003',  480,  510,  495,  525,  515,  535],
        ['SKU-H001', 2200, 2150, 2280, 2260, 2310, 2285],
        ['SKU-H002',  750,  780,  765,  800,  790,  815],
        ['SKU-H003', 1480, 1510, 1495, 1530, 1545, 1525],
        ['SKU-I001',  890,  920,  905,  940,  930,  955],
        ['SKU-I002', 2950, 2900, 3050, 3020, 3080, 3055],
        ['SKU-I003',  620,  650,  635,  670,  660,  685],
        ['SKU-J001', 1840, 1810, 1880, 1860, 1910, 1885],
        ['SKU-J002',  410,  440,  425,  455,  445,  465],
        ['SKU-J003', 3750, 3700, 3850, 3820, 3880, 3855],
    ]

    for r, row_data in enumerate(sku_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
        # Column H (Forecast Next Month) intentionally left empty

    # Column widths
    ws1.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws1.column_dimensions[col_letter].width = 14

    ws1.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Sheet 2: ForecastAccuracy                                            #
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('ForecastAccuracy')

    # Header row
    headers_fa = ['SKU', 'Forecast', 'Actual', 'Absolute Error', 'MAPE %']
    for col, h in enumerate(headers_fa, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # This month's actual demand (column C filled, columns B/D/E empty)
    actual_demand = [
        ['SKU-A001', 1558],
        ['SKU-A002',  962],
        ['SKU-A003', 2390],
        ['SKU-B001',  618],
        ['SKU-B002', 1270],
        ['SKU-B003', 3365],
        ['SKU-C001',  860],
        ['SKU-C002', 2190],
        ['SKU-C003',  520],
        ['SKU-D001', 1705],
        ['SKU-D002',  365],
        ['SKU-D003', 4275],
        ['SKU-E001', 1130],
        ['SKU-E002',  760],
        ['SKU-E003', 2910],
        ['SKU-F001', 1010],
        ['SKU-F002', 1855],
        ['SKU-F003',  270],
        ['SKU-G001', 3610],
        ['SKU-G002', 1380],
        ['SKU-G003',  545],
        ['SKU-H001', 2300],
        ['SKU-H002',  820],
        ['SKU-H003', 1545],
        ['SKU-I001',  965],
        ['SKU-I002', 3075],
        ['SKU-I003',  690],
        ['SKU-J001', 1900],
        ['SKU-J002',  475],
        ['SKU-J003', 3870],
    ]

    for r, (sku, actual) in enumerate(actual_demand, 2):
        ws2.cell(row=r, column=1, value=sku)     # SKU in col A
        # col B (Forecast) intentionally left EMPTY
        ws2.cell(row=r, column=3, value=actual)  # Actual in col C
        # col D (Absolute Error) intentionally left EMPTY
        # col E (MAPE %) intentionally left EMPTY

    # Column widths
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 12

    ws2.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  DemandHistory: 30 SKUs, 6 months of history, H column empty')
    print(f'  ForecastAccuracy: 30 SKUs, actual demand filled, B/D/E columns empty')


create_initial()
