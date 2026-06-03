"""
Reward Script: Create 3-month demand forecast using moving average and calculate MAPE
Task ID: calc_ops_supply_chain_demand_planning_056
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): DemandHistory H2:H31 contain 3-month moving average formulas
  - Component 2 (0.25): ForecastAccuracy B2:B31 contain VLOOKUP formulas pulling from DemandHistory
  - Component 3 (0.20): ForecastAccuracy D2:D31 contain absolute error formulas ABS(C-B)
  - Component 4 (0.15): ForecastAccuracy E2:E31 contain MAPE% formulas D/C
  - Component 5 (0.10): Overall MAPE summary in F2 and conditional formatting on E column
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_ops_supply_chain_demand_planning_056'


def normalize_formula(formula):
    """Normalize a formula string for comparison: uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both required sheets must exist
    if 'DemandHistory' not in wb.sheetnames or 'ForecastAccuracy' not in wb.sheetnames:
        print(f"CRITICAL: Required sheets missing. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_dh = wb['DemandHistory']
    ws_fa = wb['ForecastAccuracy']

    # -------------------------------------------------------------------
    # Component 1: DemandHistory H2:H31 — 3-month moving average formulas
    # Each cell should contain =(E{r}+F{r}+G{r})/3
    # FAILS on initial (all None), PASSES on golden (all formulas present)
    # Worth: 0.30 points
    # -------------------------------------------------------------------
    try:
        h_formula_count = 0
        h_total = 30
        h_failures = []

        for row in range(2, 32):  # rows 2..31
            cell_val = ws_dh.cell(row=row, column=8).value  # column H
            if cell_val is None:
                h_failures.append(f"H{row}: None")
                continue

            # Accept any formula that averages E, F, G columns for this row
            # Pattern: =(E{r}+F{r}+G{r})/3 or equivalents like =SUM(E{r}:G{r})/3
            normalized = normalize_formula(str(cell_val))

            # Primary pattern: =(Er+Fr+Gr)/3
            primary_pattern = f'=(E{row}+F{row}+G{row})/3'
            # Alternative SUM pattern: =SUM(Er:Gr)/3
            sum_pattern = f'=SUM(E{row}:G{row})/3'

            if (normalize_formula(primary_pattern) == normalized or
                    normalize_formula(sum_pattern) == normalized or
                    # Accept AVERAGE formula: =AVERAGE(Er:Gr) or =(Er:Gr)/3 style
                    normalize_formula(f'=AVERAGE(E{row}:G{row})') == normalized):
                h_formula_count += 1
            else:
                # Also check if it's a formula containing E, F, G columns and divides by 3
                # to handle slight variations
                if (normalized.startswith('=') and
                        f'E{row}' in normalized and
                        f'F{row}' in normalized and
                        f'G{row}' in normalized and
                        '/3' in normalized):
                    h_formula_count += 1
                else:
                    h_failures.append(f"H{row}: {repr(cell_val)}")

        if h_formula_count == h_total:
            print(f"PASS: Component 1 — All {h_total} DemandHistory H2:H31 cells have 3-month moving average formulas (0.30 pts)")
            total_score += 0.30
        elif h_formula_count >= 15:
            # Partial credit for more than half correct
            partial = round(0.30 * (h_formula_count / h_total), 2)
            print(f"PARTIAL: Component 1 — {h_formula_count}/{h_total} DemandHistory H cells have moving average formulas ({partial} pts)")
            print(f"  First failures: {h_failures[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {h_formula_count}/{h_total} DemandHistory H cells have moving average formulas")
            print(f"  First failures: {h_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: ForecastAccuracy B2:B31 — VLOOKUP formulas
    # Each cell should reference DemandHistory column 8 via VLOOKUP
    # FAILS on initial (all None), PASSES on golden (VLOOKUP formulas)
    # Worth: 0.25 points
    # -------------------------------------------------------------------
    try:
        b_formula_count = 0
        b_total = 30
        b_failures = []

        for row in range(2, 32):  # rows 2..31
            cell_val = ws_fa.cell(row=row, column=2).value  # column B
            if cell_val is None:
                b_failures.append(f"B{row}: None")
                continue

            normalized = normalize_formula(str(cell_val))

            # Must be a formula starting with = and containing VLOOKUP with DemandHistory
            # and column index 8 (the H column with forecast)
            if (normalized.startswith('=') and
                    'VLOOKUP' in normalized and
                    'DEMANDHISTORY' in normalized and
                    '8' in normalized):
                b_formula_count += 1
            else:
                b_failures.append(f"B{row}: {repr(cell_val)}")

        if b_formula_count == b_total:
            print(f"PASS: Component 2 — All {b_total} ForecastAccuracy B2:B31 cells have VLOOKUP formulas (0.25 pts)")
            total_score += 0.25
        elif b_formula_count >= 15:
            partial = round(0.25 * (b_formula_count / b_total), 2)
            print(f"PARTIAL: Component 2 — {b_formula_count}/{b_total} ForecastAccuracy B cells have VLOOKUP formulas ({partial} pts)")
            print(f"  First failures: {b_failures[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {b_formula_count}/{b_total} ForecastAccuracy B cells have VLOOKUP formulas")
            print(f"  First failures: {b_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: ForecastAccuracy D2:D31 — Absolute error formulas
    # Each cell should contain =ABS(C{r}-B{r})
    # FAILS on initial (all None), PASSES on golden
    # Worth: 0.20 points
    # -------------------------------------------------------------------
    try:
        d_formula_count = 0
        d_total = 30
        d_failures = []

        for row in range(2, 32):  # rows 2..31
            cell_val = ws_fa.cell(row=row, column=4).value  # column D
            if cell_val is None:
                d_failures.append(f"D{row}: None")
                continue

            normalized = normalize_formula(str(cell_val))
            # Pattern: =ABS(C{r}-B{r}) or =ABS(B{r}-C{r})
            pattern1 = normalize_formula(f'=ABS(C{row}-B{row})')
            pattern2 = normalize_formula(f'=ABS(B{row}-C{row})')

            if normalized == pattern1 or normalized == pattern2:
                d_formula_count += 1
            elif (normalized.startswith('=') and
                  'ABS' in normalized and
                  f'C{row}' in normalized and
                  f'B{row}' in normalized):
                d_formula_count += 1
            else:
                d_failures.append(f"D{row}: {repr(cell_val)}")

        if d_formula_count == d_total:
            print(f"PASS: Component 3 — All {d_total} ForecastAccuracy D2:D31 cells have ABS error formulas (0.20 pts)")
            total_score += 0.20
        elif d_formula_count >= 15:
            partial = round(0.20 * (d_formula_count / d_total), 2)
            print(f"PARTIAL: Component 3 — {d_formula_count}/{d_total} ForecastAccuracy D cells have ABS error formulas ({partial} pts)")
            print(f"  First failures: {d_failures[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {d_formula_count}/{d_total} ForecastAccuracy D cells have ABS error formulas")
            print(f"  First failures: {d_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: ForecastAccuracy E2:E31 — MAPE percentage formulas
    # Each cell should contain =D{r}/C{r} formatted as percentage
    # FAILS on initial (all None), PASSES on golden
    # Worth: 0.15 points
    # -------------------------------------------------------------------
    try:
        e_formula_count = 0
        e_total = 30
        e_failures = []

        for row in range(2, 32):  # rows 2..31
            cell_val = ws_fa.cell(row=row, column=5).value  # column E
            if cell_val is None:
                e_failures.append(f"E{row}: None")
                continue

            normalized = normalize_formula(str(cell_val))
            # Pattern: =D{r}/C{r}
            pattern = normalize_formula(f'=D{row}/C{row}')

            if normalized == pattern:
                e_formula_count += 1
            elif (normalized.startswith('=') and
                  f'D{row}' in normalized and
                  f'C{row}' in normalized and
                  '/' in normalized):
                e_formula_count += 1
            else:
                e_failures.append(f"E{row}: {repr(cell_val)}")

        if e_formula_count == e_total:
            print(f"PASS: Component 4 — All {e_total} ForecastAccuracy E2:E31 cells have MAPE% formulas (0.15 pts)")
            total_score += 0.15
        elif e_formula_count >= 15:
            partial = round(0.15 * (e_formula_count / e_total), 2)
            print(f"PARTIAL: Component 4 — {e_formula_count}/{e_total} ForecastAccuracy E cells have MAPE formulas ({partial} pts)")
            print(f"  First failures: {e_failures[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {e_formula_count}/{e_total} ForecastAccuracy E cells have MAPE formulas")
            print(f"  First failures: {e_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------
    # Component 5: Overall MAPE summary (F2) AND conditional formatting on E column
    # F2 should contain =AVERAGE(E2:E31) as the overall MAPE summary
    # E column should have conditional formatting (red >20%, green <=10%)
    # FAILS on initial (F2 is None, no CF), PASSES on golden
    # Worth: 0.10 points (0.05 for F2 summary, 0.05 for CF)
    # -------------------------------------------------------------------
    try:
        sub_score = 0.0

        # Sub-check 5a: F2 overall MAPE summary (0.05)
        f2_val = ws_fa.cell(row=2, column=6).value
        if f2_val is not None:
            f2_normalized = normalize_formula(str(f2_val))
            expected_f2 = normalize_formula('=AVERAGE(E2:E31)')
            if expected_f2 == f2_normalized or (
                    f2_normalized.startswith('=') and
                    'AVERAGE' in f2_normalized and
                    'E2' in f2_normalized and
                    'E31' in f2_normalized):
                print(f"PASS: Component 5a — F2 contains overall MAPE summary formula: {repr(f2_val)} (0.05 pts)")
                sub_score += 0.05
            else:
                print(f"FAIL: Component 5a — F2 does not contain expected AVERAGE formula, found: {repr(f2_val)}")
        else:
            print(f"FAIL: Component 5a — F2 is empty (expected AVERAGE(E2:E31) for overall MAPE)")

        # Sub-check 5b: Conditional formatting exists on E column (0.05)
        cf_on_e = False
        for cf_range in ws_fa.conditional_formatting:
            cf_str = str(cf_range)
            if 'E' in cf_str and cf_str != 'E1':
                # Check if range covers E2:E31 or part of it
                cf_rules = ws_fa.conditional_formatting[cf_range]
                if len(list(cf_rules)) > 0:
                    cf_on_e = True
                    print(f"PASS: Component 5b — Conditional formatting found on range {cf_str} (0.05 pts)")
                    break

        if cf_on_e:
            sub_score += 0.05
        else:
            print(f"FAIL: Component 5b — No conditional formatting found on E column (expected red >20%, green <=10%)")

        if sub_score > 0:
            print(f"Component 5 total: {sub_score} pts")
            total_score += sub_score

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
