import subprocess
import sys
import time

# Persistence hook: save the file via Ctrl+S before verification
try:
    subprocess.Popen(
        ["xdotool", "key", "ctrl+s"],
        env={"DISPLAY": ":0", "PATH": "/usr/bin:/bin"},
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    time.sleep(2)
except Exception:
    pass

import openpyxl

FILE_PATH = "/home/user/calc_gg5_020.xlsx"


def compute_reward() -> float:
    score = 0.0

    try:
        wb = openpyxl.load_workbook(FILE_PATH)
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return 0.0

    # Check that the Actuals sheet exists
    if "Actuals" not in wb.sheetnames:
        print("Sheet 'Actuals' not found")
        return 0.0

    ws = wb["Actuals"]

    # Component 1: E2:E50 cells are unlocked (0.35 pts)
    # Check a representative sample of cells in the range
    sample_cells = ["E2", "E5", "E10", "E20", "E30", "E40", "E50"]
    unlocked_count = 0
    for coord in sample_cells:
        try:
            if ws[coord].protection.locked is False:
                unlocked_count += 1
        except Exception:
            pass

    if unlocked_count == len(sample_cells):
        score += 0.35
        print(f"Component 1 (E2:E50 unlocked): 0.35 — all {unlocked_count}/{len(sample_cells)} sample cells unlocked")
    elif unlocked_count > 0:
        partial = 0.35 * (unlocked_count / len(sample_cells))
        score += partial
        print(f"Component 1 (E2:E50 unlocked): {partial:.2f} — {unlocked_count}/{len(sample_cells)} sample cells unlocked")
    else:
        print(f"Component 1 (E2:E50 unlocked): 0.00 — no cells unlocked")

    # Component 2: Sheet protection enabled with password (0.30 pts)
    prot_score = 0.0
    if ws.protection.sheet is True:
        prot_score += 0.15
        print("Component 2a (sheet protected): 0.15")
    else:
        print("Component 2a (sheet protected): 0.00 — sheet not protected")

    if ws.protection.password is not None and ws.protection.password != "":
        prot_score += 0.15
        print("Component 2b (password set): 0.15")
    else:
        print("Component 2b (password set): 0.00 — no password")

    score += prot_score

    # Component 3: Protection allows selectUnlockedCells and formatCells (0.20 pts)
    # Only meaningful when sheet protection is enabled
    # In openpyxl: False = allowed, True = blocked
    sheet_is_protected = ws.protection.sheet is True
    comp3_score = 0.0
    if sheet_is_protected:
        if ws.protection.selectUnlockedCells is False:
            comp3_score += 0.10
            print("Component 3a (selectUnlockedCells allowed): 0.10")
        else:
            print("Component 3a (selectUnlockedCells allowed): 0.00 — blocked")

        if ws.protection.formatCells is False:
            comp3_score += 0.10
            print("Component 3b (formatCells allowed): 0.10")
        else:
            print("Component 3b (formatCells allowed): 0.00 — blocked")
    else:
        print("Component 3 (protection settings): 0.00 — sheet not protected, settings irrelevant")

    score += comp3_score

    # Component 4: Other cells remain locked while E column is unlocked (0.15 pts)
    # Only meaningful when sheet protection is enabled AND E cells are unlocked
    # This verifies the selective unlock pattern
    if sheet_is_protected and unlocked_count > 0:
        locked_check_cells = ["A2", "B2", "C2", "D2", "F2", "A10", "B10"]
        locked_count = 0
        for coord in locked_check_cells:
            try:
                if ws[coord].protection.locked is not False:
                    locked_count += 1
            except Exception:
                locked_count += 1  # default is locked

        if locked_count == len(locked_check_cells):
            score += 0.15
            print(f"Component 4 (other cells locked): 0.15 — all {locked_count}/{len(locked_check_cells)} cells locked")
        elif locked_count > 0:
            partial = 0.15 * (locked_count / len(locked_check_cells))
            score += partial
            print(f"Component 4 (other cells locked): {partial:.2f} — {locked_count}/{len(locked_check_cells)} locked")
        else:
            print(f"Component 4 (other cells locked): 0.00")
    else:
        print("Component 4 (other cells locked): 0.00 — prerequisite not met")

    # Round to avoid floating point issues
    score = round(score, 2)
    print(f"\nTotal score: {score}")
    return score


reward = compute_reward()
print(f"reward: {reward}")
