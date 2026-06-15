"""initial_setup.py - Create project_costs.xlsx with project data, no protection."""

import subprocess
subprocess.check_call(['pip3', 'install', 'openpyxl'], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)

import os
import shlex
import subprocess
import time
import datetime
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_020'
OUTPUT_PATH = os.path.join(WORKDIR, f'{TASK_ID}.xlsx')

# ---------- Build workbook ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Actuals'

# Headers
headers = ['Project ID', 'Name', 'Budgeted Cost', 'Start Date', 'Actual Cost', 'Variance']
header_font = Font(bold=True, size=11, name='Calibri')
header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center')
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = border

# Column widths
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 36
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 14
ws.column_dimensions['E'].width = 16
ws.column_dimensions['F'].width = 16

# Project data - 49 rows (rows 2-50)
project_names = [
    'Website Redesign Phase 1', 'ERP System Migration', 'Mobile App Development',
    'Cloud Infrastructure Setup', 'Data Warehouse Build', 'CRM Integration Project',
    'Security Audit 2024', 'Network Upgrade East', 'Office Renovation Floor 3',
    'HR Portal Revamp', 'Supply Chain Optimization', 'Customer Analytics Dashboard',
    'Payment Gateway Update', 'Inventory Management System', 'Employee Training Portal',
    'Marketing Automation Setup', 'Disaster Recovery Plan', 'API Gateway Development',
    'DevOps Pipeline Setup', 'Quality Assurance Framework', 'Vendor Management Portal',
    'Business Intelligence Tool', 'Compliance Tracking System', 'Digital Signage Network',
    'Warehouse Robotics Pilot', 'Fleet Management System', 'Energy Monitoring Platform',
    'Social Media Integration', 'Chatbot Implementation', 'Document Management System',
    'Video Conferencing Upgrade', 'Cybersecurity Training', 'Database Optimization',
    'Load Balancer Configuration', 'SSO Implementation', 'Backup System Overhaul',
    'Print Management Solution', 'Asset Tracking System', 'Visitor Management App',
    'Parking System Upgrade', 'HVAC Monitoring IoT', 'Badge Access Expansion',
    'Cafeteria POS Upgrade', 'Fire Safety Compliance', 'Elevator Modernization',
    'Landscaping Renovation', 'Signage Refresh Program', 'Reception Area Remodel',
    'Server Room Cooling',
]

random.seed(42)

currency_fmt = '$#,##0.00'
date_fmt = 'yyyy-mm-dd'

for i in range(49):
    row = i + 2
    pid = f'PRJ-{2024000 + i + 1:07d}'
    name = project_names[i]
    budgeted = round(random.uniform(15000, 500000), 2)
    start = datetime.date(2024, 1, 1) + datetime.timedelta(days=random.randint(0, 300))
    actual = round(random.uniform(budgeted * 0.7, budgeted * 1.3), 2)
    variance = round(budgeted - actual, 2)

    ws.cell(row=row, column=1, value=pid).border = border
    ws.cell(row=row, column=2, value=name).border = border
    c_cell = ws.cell(row=row, column=3, value=budgeted)
    c_cell.number_format = currency_fmt
    c_cell.border = border
    d_cell = ws.cell(row=row, column=4, value=start)
    d_cell.number_format = date_fmt
    d_cell.border = border
    e_cell = ws.cell(row=row, column=5, value=actual)
    e_cell.number_format = currency_fmt
    e_cell.border = border
    f_cell = ws.cell(row=row, column=6, value=variance)
    f_cell.number_format = currency_fmt
    f_cell.border = border

# Freeze header row
ws.freeze_panes = 'A2'

# Save
wb.save(OUTPUT_PATH)
print(f'Created {OUTPUT_PATH}')

# ---------- Launch LibreOffice Calc ----------
env = os.environ.copy()
env['DISPLAY'] = ':0'
subprocess.Popen(
    shlex.split(f'libreoffice --calc "{OUTPUT_PATH}"'),
    stdout=subprocess.DEVNULL,
    stderr=subprocess.DEVNULL,
    env=env,
)
time.sleep(2)
print('LibreOffice Calc launched.')
