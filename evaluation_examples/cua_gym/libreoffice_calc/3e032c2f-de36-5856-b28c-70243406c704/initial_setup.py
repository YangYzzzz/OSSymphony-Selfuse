"""
Initial Setup: Create Orders spreadsheet for pivot table task
Task ID: calc_pivot_032
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orders'

    # Headers
    headers = ['OrderID', 'CustomerName', 'Segment', 'OrderValue', 'Items', 'ShipDate']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    # Segment distribution:
    # Consumer: 200 rows, sum=68000, avg=340
    # Corporate: 140 rows, sum=85000, avg=607.14...
    # Enterprise: 100 rows, sum=112000, avg=1120
    # Government: 80 rows, sum=55000, avg=687.5
    # Total: 520 rows

    segments_config = [
        ('Consumer', 200, 68000),
        ('Corporate', 140, 85000),
        ('Enterprise', 100, 112000),
        ('Government', 80, 55000),
    ]

    first_names = [
        'Sarah', 'Marcus', 'Elena', 'David', 'Priya', 'James', 'Aiko', 'Carlos',
        'Fatima', 'Thomas', 'Mei', 'Robert', 'Anna', 'Kwame', 'Sofia', 'Daniel',
        'Rachel', 'Ibrahim', 'Linda', 'Yuki', 'Oliver', 'Natasha', 'Ahmed',
        'Grace', 'Patrick', 'Maya', 'Victor', 'Hannah', 'Leo', 'Diana',
        'Samuel', 'Zara', 'Michael', 'Emma', 'Kevin', 'Laura', 'Raj', 'Claire',
        'Oscar', 'Julia', 'Nathan', 'Isla', 'Felix', 'Nina', 'Hugo', 'Chloe',
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Gupta', 'O\'Brien', 'Tanaka',
        'Rodriguez', 'Al-Rashid', 'Fischer', 'Zhang', 'Thompson', 'Kowalski',
        'Mensah', 'Garcia', 'Park', 'Goldberg', 'Hassan', 'Kumar', 'Nakamura',
        'Andersson', 'Morales', 'Kim', 'Brown', 'Dubois', 'Patel', 'Wilson',
        'Martinez', 'Taylor', 'Lee', 'Harris', 'Clark', 'Lewis', 'Walker',
        'Young', 'Allen', 'King', 'Wright', 'Scott', 'Adams', 'Baker', 'Hill',
    ]

    base_year = 2024
    months = list(range(1, 13))
    days_in_month = {1:31, 2:28, 3:31, 4:30, 5:31, 6:30, 7:31, 8:31, 9:30, 10:31, 11:30, 12:31}

    def random_date():
        m = random.choice(months)
        d = random.randint(1, days_in_month[m])
        y = random.choice([2024, 2025])
        return f'{y}-{m:02d}-{d:02d}'

    def generate_order_values(count, target_sum):
        """Generate `count` random positive values that sum to target_sum."""
        # Generate random proportions
        raw = [random.random() + 0.1 for _ in range(count)]
        total_raw = sum(raw)
        # Scale to target sum, round to 2 decimals
        values = [round((r / total_raw) * target_sum, 2) for r in raw]
        # Fix rounding difference on last value
        diff = round(target_sum - sum(values), 2)
        values[-1] = round(values[-1] + diff, 2)
        return values

    all_rows = []
    for segment, count, target_sum in segments_config:
        order_values = generate_order_values(count, target_sum)
        for i in range(count):
            name = f'{random.choice(first_names)} {random.choice(last_names)}'
            items = random.randint(1, 15)
            ship_date = random_date()
            all_rows.append((name, segment, order_values[i], items, ship_date))

    # Shuffle to mix segments
    random.shuffle(all_rows)

    # Write data rows
    for idx, (name, segment, order_val, items, ship_date) in enumerate(all_rows, 2):
        ws.cell(row=idx, column=1, value=idx - 1)  # OrderID
        ws.cell(row=idx, column=2, value=name)
        ws.cell(row=idx, column=3, value=segment)
        ws.cell(row=idx, column=4, value=order_val)
        ws.cell(row=idx, column=5, value=items)
        ws.cell(row=idx, column=6, value=ship_date)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 14

    # Number format for OrderValue
    for r in range(2, len(all_rows) + 2):
        ws.cell(row=r, column=4).number_format = '#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total data rows: {len(all_rows)}')

    # Verify sums
    from collections import defaultdict
    sums = defaultdict(float)
    counts = defaultdict(int)
    for name, segment, order_val, items, ship_date in all_rows:
        sums[segment] += order_val
        counts[segment] += 1
    for seg in ['Consumer', 'Corporate', 'Enterprise', 'Government']:
        avg = sums[seg] / counts[seg] if counts[seg] > 0 else 0
        print(f'  {seg}: count={counts[seg]}, sum={sums[seg]:.2f}, avg={avg:.2f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
