"""
Reward Script: Pivot table with SUM and AVG of OrderValue by Segment
Task ID: calc_pivot_032
Domain: libreoffice_calc
Scoring:
  - Component 1: Pivot table sheet exists (0.2)
  - Component 2: Correct column structure with Segment + SUM + AVG headers (0.15)
  - Component 3: All 4 segments present as rows (0.15)
  - Component 4: SUM of OrderValue values correct for all segments (0.25)
  - Component 5: AVG of OrderValue values correct for all segments (0.25)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_032'

# Expected ground truth values from task context
EXPECTED_DATA = {
    'Consumer':   {'sum': 68000,  'avg': 340.0},
    'Corporate':  {'sum': 85000,  'avg': 607.14},
    'Enterprise': {'sum': 112000, 'avg': 1120.0},
    'Government': {'sum': 55000,  'avg': 687.5},
}

# Tolerance for numeric comparisons
SUM_TOLERANCE = 500    # sums are large numbers, allow small rounding
AVG_TOLERANCE = 5.0    # averages allow small rounding


def find_pivot_sheet(wb):
    """Find a sheet that looks like a pivot table (not 'Orders')."""
    for name in wb.sheetnames:
        if name.lower() != 'orders':
            return name
    return None


def normalize(s):
    """Lowercase, strip whitespace and special chars for fuzzy matching."""
    if s is None:
        return ''
    return re.sub(r'[^a-z0-9]', '', str(s).lower().strip())


def has_keyword(text, keywords):
    """Check if normalized text contains any of the keywords."""
    n = normalize(text)
    return any(kw in n for kw in keywords)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot table sheet exists (0.2 points)
    # This checks that a NEW sheet was created (beyond just 'Orders')
    try:
        pivot_sheet_name = find_pivot_sheet(wb)
        if pivot_sheet_name is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_sheet_name}' (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — No pivot table sheet found. Only sheets: {wb.sheetnames}")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb[pivot_sheet_name]

    # Build a map of data from the pivot sheet
    # Scan all rows to find segment rows and header row
    # We need to be flexible about exact layout

    # First, find the header row (containing references to sum/average and segment)
    header_row = None
    sum_col = None
    avg_col = None
    segment_col = None

    try:
        for r in range(1, min(ws.max_row + 1, 20)):
            row_vals = []
            for c in range(1, min(ws.max_column + 1, 10)):
                row_vals.append((c, ws.cell(r, c).value))

            # Check if this row has headers
            found_segment = False
            found_sum = False
            found_avg = False
            for c, v in row_vals:
                if v is None:
                    continue
                nv = normalize(v)
                if 'segment' in nv:
                    segment_col = c
                    found_segment = True
                elif has_keyword(v, ['sum', 'total']) and has_keyword(v, ['order', 'value']):
                    sum_col = c
                    found_sum = True
                elif has_keyword(v, ['average', 'avg', 'mean']) and has_keyword(v, ['order', 'value']):
                    avg_col = c
                    found_avg = True

            if found_segment and (found_sum or found_avg):
                header_row = r
                break

        # Component 2: Correct column structure (0.15 points)
        if header_row is not None and sum_col is not None and avg_col is not None and segment_col is not None:
            print(f"PASS: Component 2 — Headers found at row {header_row}: segment col={segment_col}, sum col={sum_col}, avg col={avg_col} (0.15 pts)")
            total_score += 0.15
        elif header_row is not None:
            # Partial: found some headers but not all
            print(f"FAIL: Component 2 — Incomplete headers. segment_col={segment_col}, sum_col={sum_col}, avg_col={avg_col}")
        else:
            print(f"FAIL: Component 2 — No header row with Segment + SUM/AVG columns found")
            # Try fallback: maybe columns are positional (A=Segment, B=Sum, C=Avg)
            # Check if row 1 has some content
            if ws.cell(1, 1).value is not None:
                segment_col = 1
                sum_col = 2
                avg_col = 3
                header_row = 1
                print(f"  INFO: Falling back to positional columns (A=Segment, B=Sum, C=Avg)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    if header_row is None or segment_col is None:
        print(f"REWARD: {total_score}")
        return total_score

    # Read data rows from pivot sheet
    pivot_data = {}
    try:
        for r in range(header_row + 1, ws.max_row + 1):
            seg_val = ws.cell(r, segment_col).value
            if seg_val is None:
                continue
            seg_name = str(seg_val).strip()
            # Skip grand total row for segment matching
            if normalize(seg_name) in ['grandtotal', 'total']:
                continue
            sum_val = ws.cell(r, sum_col).value if sum_col else None
            avg_val = ws.cell(r, avg_col).value if avg_col else None
            pivot_data[seg_name] = {'sum': sum_val, 'avg': avg_val}
    except Exception as e:
        print(f"ERROR: Reading pivot data — {e}")

    # Component 3: All 4 segments present (0.15 points)
    try:
        expected_segments = set(EXPECTED_DATA.keys())
        found_segments = set()
        for seg in pivot_data:
            for exp_seg in expected_segments:
                if normalize(seg) == normalize(exp_seg):
                    found_segments.add(exp_seg)

        if found_segments == expected_segments:
            print(f"PASS: Component 3 — All 4 segments found: {sorted(found_segments)} (0.15 pts)")
            total_score += 0.15
        else:
            missing = expected_segments - found_segments
            print(f"FAIL: Component 3 — Missing segments: {missing}. Found: {sorted(pivot_data.keys())}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Helper to match segment names flexibly
    def match_segment(name):
        for exp in EXPECTED_DATA:
            if normalize(name) == normalize(exp):
                return exp
        return None

    # Component 4: SUM values correct (0.25 points)
    # Award partial credit per segment (0.0625 each)
    try:
        sum_correct = 0
        for seg_name, data in pivot_data.items():
            exp_seg = match_segment(seg_name)
            if exp_seg is None:
                continue
            expected_sum = EXPECTED_DATA[exp_seg]['sum']
            actual_sum = data['sum']
            if actual_sum is not None:
                try:
                    if abs(float(actual_sum) - expected_sum) <= SUM_TOLERANCE:
                        sum_correct += 1
                        print(f"  SUM OK: {exp_seg} = {actual_sum} (expected ~{expected_sum})")
                    else:
                        print(f"  SUM MISMATCH: {exp_seg} = {actual_sum} (expected ~{expected_sum})")
                except (ValueError, TypeError):
                    print(f"  SUM ERROR: {exp_seg} = {actual_sum} (not numeric)")
            else:
                print(f"  SUM MISSING: {exp_seg} = None")

        sum_score = (sum_correct / 4) * 0.25
        if sum_correct == 4:
            print(f"PASS: Component 4 — All 4 SUM values correct (0.25 pts)")
        elif sum_correct > 0:
            print(f"PARTIAL: Component 4 — {sum_correct}/4 SUM values correct ({sum_score:.3f} pts)")
        else:
            print(f"FAIL: Component 4 — No SUM values correct")
        total_score += sum_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: AVG values correct (0.25 points)
    # Award partial credit per segment (0.0625 each)
    try:
        avg_correct = 0
        for seg_name, data in pivot_data.items():
            exp_seg = match_segment(seg_name)
            if exp_seg is None:
                continue
            expected_avg = EXPECTED_DATA[exp_seg]['avg']
            actual_avg = data['avg']
            if actual_avg is not None:
                try:
                    if abs(float(actual_avg) - expected_avg) <= AVG_TOLERANCE:
                        avg_correct += 1
                        print(f"  AVG OK: {exp_seg} = {actual_avg} (expected ~{expected_avg})")
                    else:
                        print(f"  AVG MISMATCH: {exp_seg} = {actual_avg} (expected ~{expected_avg})")
                except (ValueError, TypeError):
                    print(f"  AVG ERROR: {exp_seg} = {actual_avg} (not numeric)")
            else:
                print(f"  AVG MISSING: {exp_seg} = None")

        avg_score = (avg_correct / 4) * 0.25
        if avg_correct == 4:
            print(f"PASS: Component 5 — All 4 AVG values correct (0.25 pts)")
        elif avg_correct > 0:
            print(f"PARTIAL: Component 5 — {avg_correct}/4 AVG values correct ({avg_score:.3f} pts)")
        else:
            print(f"FAIL: Component 5 — No AVG values correct")
        total_score += avg_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice (save any unsaved changes)
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
