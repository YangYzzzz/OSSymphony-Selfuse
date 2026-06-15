"""
Reward Script: Build a sales leaderboard for the monthly all-hands meeting.
Task ID: calc_sales_report_leaderboard_047
Domain: libreoffice_calc
Scoring:
  Component 1: RANK formula in A2:A16 — 0.30 pts
  Component 2: Attainment formula (=Cx/Dx) in E2:E16 — 0.20 pts
  Component 3: Days Remaining formula (=$H$1-TODAY()) in F2:F16 — 0.15 pts
  Component 4: Data sorted by Revenue (C) descending — 0.15 pts
  Component 5: Conditional formatting: gold(rank=1), silver(rank=2), bronze(rank=3) — 0.15 pts
  Component 6: Currency number format on C and D columns — 0.05 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_report_leaderboard_047'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Leaderboard' not in wb.sheetnames:
        print("FAIL: 'Leaderboard' sheet not found")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Leaderboard']

    # Component 1: RANK formula in A2:A16 (0.30 points)
    # Each row in A2:A16 should have a RANK formula referencing $C$2:$C$16
    try:
        rank_formula_count = 0
        rank_formula_correct = 0
        for row in range(2, 17):
            val = ws.cell(row=row, column=1).value
            if val is not None and isinstance(val, str) and val.upper().startswith('=RANK('):
                rank_formula_count += 1
                # Check that it references the full range $C$2:$C$16 with absolute references
                # and uses descending order (0)
                val_upper = val.upper().replace(' ', '')
                if '$C$2:$C$16' in val_upper and ',0)' in val_upper:
                    rank_formula_correct += 1

        if rank_formula_count == 15 and rank_formula_correct == 15:
            print(f"PASS: Component 1 — All 15 RANK formulas present with correct range and order (0.30 pts)")
            total_score += 0.30
        elif rank_formula_count == 15:
            print(f"PASS (partial): Component 1 — All 15 RANK formulas present but {15 - rank_formula_correct} have incorrect range/order (0.15 pts)")
            total_score += 0.15
        elif rank_formula_count > 0:
            print(f"FAIL: Component 1 — Only {rank_formula_count}/15 RANK formulas found in column A")
        else:
            print(f"FAIL: Component 1 — No RANK formulas found in A2:A16 (all cells are empty or non-formula)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Attainment formula (=Cx/Dx) in E2:E16 (0.20 points)
    # Each row in E2:E16 should have a division formula: =C2/D2 pattern
    try:
        attain_formula_count = 0
        for row in range(2, 17):
            val = ws.cell(row=row, column=5).value
            if val is not None and isinstance(val, str):
                val_stripped = val.upper().replace(' ', '')
                # Accept =Cx/Dx patterns (e.g., =C2/D2)
                if val_stripped.startswith('=C') and '/D' in val_stripped:
                    attain_formula_count += 1

        if attain_formula_count == 15:
            print(f"PASS: Component 2 — All 15 Attainment formulas (=Cx/Dx) present in E2:E16 (0.20 pts)")
            total_score += 0.20
        elif attain_formula_count > 0:
            print(f"FAIL (partial): Component 2 — Only {attain_formula_count}/15 Attainment formulas found in column E")
        else:
            print(f"FAIL: Component 2 — No Attainment formulas found in E2:E16")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Days Remaining formula (=$H$1-TODAY()) in F2:F16 (0.15 points)
    try:
        days_formula_count = 0
        for row in range(2, 17):
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str):
                val_stripped = val.upper().replace(' ', '')
                # Accept =$H$1-TODAY() pattern
                if '$H$1-TODAY()' in val_stripped or 'H1-TODAY()' in val_stripped:
                    days_formula_count += 1

        if days_formula_count == 15:
            print(f"PASS: Component 3 — All 15 Days Remaining formulas (=$H$1-TODAY()) present in F2:F16 (0.15 pts)")
            total_score += 0.15
        elif days_formula_count > 0:
            print(f"FAIL (partial): Component 3 — Only {days_formula_count}/15 Days Remaining formulas found in column F")
        else:
            print(f"FAIL: Component 3 — No Days Remaining formulas found in F2:F16")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data sorted by Revenue (column C) descending (0.15 points)
    # Check that C2:C16 values are in non-increasing order
    try:
        revenues = []
        for row in range(2, 17):
            val = ws.cell(row=row, column=3).value
            if val is not None:
                try:
                    revenues.append(float(val))
                except (ValueError, TypeError):
                    revenues.append(None)

        if len(revenues) == 15 and all(r is not None for r in revenues):
            is_sorted_desc = all(revenues[i] >= revenues[i+1] for i in range(len(revenues)-1))
            if is_sorted_desc:
                print(f"PASS: Component 4 — Data sorted by Revenue descending (top={revenues[0]}, bottom={revenues[-1]}) (0.15 pts)")
                total_score += 0.15
            else:
                # Check if at least partially sorted (first element is the highest)
                max_rev = max(revenues)
                if revenues[0] == max_rev:
                    print(f"FAIL (partial): Component 4 — Revenue column not fully sorted descending (first is max but order not maintained)")
                else:
                    print(f"FAIL: Component 4 — Revenue column not sorted descending (first={revenues[0]}, max={max_rev})")
        else:
            print(f"FAIL: Component 4 — Could not read 15 revenue values from C2:C16 (got {len(revenues)} values)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting with gold/silver/bronze colors (0.15 points)
    # Check for 3 conditional formatting rules on the leaderboard range
    try:
        cf_rules = ws.conditional_formatting
        cf_list = list(cf_rules)

        # Gold: FFD700-like colors (gold is FFFFD700 or similar)
        GOLD_COLORS = ['FFFFD700', 'FFD700', 'FFF5C518', 'FFFFF0A0']
        # Silver: C0C0C0-like colors
        SILVER_COLORS = ['FFC0C0C0', 'C0C0C0', 'FFD3D3D3', 'FFBEBEBE']
        # Bronze/Copper: CD7F32-like colors
        BRONZE_COLORS = ['FFCD7F32', 'CD7F32', 'FFB87333', 'FFC87137']

        # Count matching rules per color category using list comprehension approach
        gold_matches = 0
        silver_matches = 0
        bronze_matches = 0

        for cf in cf_list:
            for rule in cf.rules:
                try:
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        fill_color = rule.dxf.fill.fgColor.rgb
                        formula_str = str(rule.formula) if hasattr(rule, 'formula') else ''
                        # Count gold rules: color is gold AND formula references rank=1
                        if fill_color in GOLD_COLORS and ('=1' in formula_str or fill_color in GOLD_COLORS):
                            gold_matches += 1
                        # Count silver rules: color is silver AND formula references rank=2
                        if fill_color in SILVER_COLORS and ('=2' in formula_str or fill_color in SILVER_COLORS):
                            silver_matches += 1
                        # Count bronze rules: color is bronze AND formula references rank=3
                        if fill_color in BRONZE_COLORS and ('=3' in formula_str or fill_color in BRONZE_COLORS):
                            bronze_matches += 1
                except Exception:
                    pass

        rules_found = sum([gold_matches > 0, silver_matches > 0, bronze_matches > 0])
        total_cf_rules = sum(len(cf.rules) for cf in cf_list)

        if gold_matches > 0 and silver_matches > 0 and bronze_matches > 0:
            print(f"PASS: Component 5 — All 3 conditional formatting rules found (gold, silver, bronze) (0.15 pts)")
            total_score += 0.15
        elif rules_found >= 2:
            print(f"PASS (partial): Component 5 — {rules_found}/3 color rules found (0.07 pts)")
            total_score += 0.07
        elif total_cf_rules > 0:
            print(f"FAIL (partial): Component 5 — {total_cf_rules} conditional formatting rule(s) exist but gold/silver/bronze not all detected ({rules_found}/3)")
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules found in Leaderboard sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Currency format on C and D columns (0.05 points)
    # Check that cells in C2:C16 and D2:D16 have a currency-like format
    try:
        currency_formats_c = 0
        currency_formats_d = 0

        for row in range(2, 17):
            c_fmt = ws.cell(row=row, column=3).number_format
            d_fmt = ws.cell(row=row, column=4).number_format
            # Currency formats typically contain $ or currency symbol
            if c_fmt and ('$' in c_fmt or '#,##0' in c_fmt) and c_fmt != 'General':
                currency_formats_c += 1
            if d_fmt and ('$' in d_fmt or '#,##0' in d_fmt) and d_fmt != 'General':
                currency_formats_d += 1

        if currency_formats_c >= 10 and currency_formats_d >= 10:
            print(f"PASS: Component 6 — Currency format applied to C ({currency_formats_c}/15) and D ({currency_formats_d}/15) columns (0.05 pts)")
            total_score += 0.05
        elif currency_formats_c >= 5 or currency_formats_d >= 5:
            print(f"FAIL (partial): Component 6 — Currency format partially applied: C={currency_formats_c}/15, D={currency_formats_d}/15")
        else:
            print(f"FAIL: Component 6 — No currency format on C column ({currency_formats_c}/15) or D column ({currency_formats_d}/15)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
