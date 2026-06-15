"""
Initial Setup: Sales Leaderboard for Monthly All-Hands Meeting
Task ID: calc_sales_report_leaderboard_047
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_report_leaderboard_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Leaderboard ---
    ws = wb.active
    ws.title = 'Leaderboard'

    # Headers in row 1
    headers = ['Rank', 'Rep Name', 'Revenue', 'Quota', 'Attainment', 'Days Remaining']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # End-of-month date stored in H1
    ws['H1'] = date(2025, 3, 31)
    ws['H1'].number_format = 'yyyy-mm-dd'
    ws['G1'] = 'Month End'

    # 15 sales reps — realistic names, revenues NOT sorted so the agent has to sort them
    # Columns: Rep Name (B), Revenue (C), Quota (D)
    # Columns A (Rank), E (Attainment), F (Days Remaining) are intentionally EMPTY
    reps_data = [
        # (Rep Name, Revenue, Quota)
        ('Priya Mehta',        183500, 175000),
        ('Daniel Nguyen',      142800, 160000),
        ('Sofia Reyes',        221400, 200000),
        ('James Whitfield',     98700, 140000),
        ('Anika Patel',        165200, 155000),
        ('Carlos Dominguez',   310500, 280000),
        ('Rachel Kim',         255800, 240000),
        ('Trevor Okafor',       75300, 120000),
        ('Mei-Ling Huang',     198600, 190000),
        ('Brandon Schultz',    134100, 150000),
        ('Valentina Cruz',     289700, 270000),
        ('Liam O\'Brien',       61200, 100000),
        ('Nadia Volkov',       177400, 165000),
        ('Marcus Adeyemi',     245900, 230000),
        ('Chloe Bergmann',     118600, 135000),
    ]

    for row_idx, (name, revenue, quota) in enumerate(reps_data, 2):
        # A (Rank) — intentionally left empty; agent will add RANK formula
        ws.cell(row=row_idx, column=2, value=name)    # B: Rep Name
        c_cell = ws.cell(row=row_idx, column=3, value=revenue)   # C: Revenue
        d_cell = ws.cell(row=row_idx, column=4, value=quota)     # D: Quota
        # E (Attainment) — intentionally left empty; agent will add formula
        # F (Days Remaining) — intentionally left empty; agent will add formula

    # Set column widths for readability
    ws.column_dimensions['A'].width = 8   # Rank
    ws.column_dimensions['B'].width = 22  # Rep Name
    ws.column_dimensions['C'].width = 14  # Revenue
    ws.column_dimensions['D'].width = 14  # Quota
    ws.column_dimensions['E'].width = 14  # Attainment
    ws.column_dimensions['F'].width = 16  # Days Remaining
    ws.column_dimensions['G'].width = 12  # Month End label
    ws.column_dimensions['H'].width = 14  # End of month date

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: Leaderboard')
    print('  - Row 1: Headers (Rank, Rep Name, Revenue, Quota, Attainment, Days Remaining)')
    print('  - Rows 2-16: 15 reps with Rep Name, Revenue, Quota filled')
    print('  - Column A (Rank): EMPTY (agent will add RANK formula)')
    print('  - Column E (Attainment): EMPTY (agent will add formula)')
    print('  - Column F (Days Remaining): EMPTY (agent will add formula)')
    print('  - H1: end-of-month date (2025-03-31)')
    print('  - No conditional formatting on data rows')
    print('  - No currency format on C/D (agent will add)')


create_initial()
