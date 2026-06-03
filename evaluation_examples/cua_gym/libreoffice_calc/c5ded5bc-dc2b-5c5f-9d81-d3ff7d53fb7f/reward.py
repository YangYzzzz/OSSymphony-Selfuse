"""
Reward Script: Use VLOOKUP with approximate match to assign letter grades to students
Task ID: osworld_calc_vlookup_grade_lookup_001
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): VLOOKUP formulas present in column D (rows 2-30)
  Component 2 (0.3 pts): Formulas use approximate match (4th argument TRUE/1)
  Component 3 (0.2 pts): Reference table F2:G6 is intact with correct grade scale
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_grade_lookup_001'

# Expected grade scale in reference table F2:G6
EXPECTED_GRADE_SCALE = {
    0: 'F',
    60: 'D',
    70: 'C',
    80: 'B',
    90: 'A',
}

# Data rows with VLOOKUP formulas (29 students)
DATA_ROWS = list(range(2, 31))  # rows 2 through 30


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires adding VLOOKUP formulas with approximate match
    in column D (rows 2-30) to assign letter grades based on scores in
    column C and a reference table in F2:G6.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet "Students" must exist
    if 'Students' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Students' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Students']

    # -------------------------------------------------------------------------
    # Component 1: VLOOKUP formulas present in column D, rows 2-30 (0.5 points)
    # In initial_env, column D rows 2-30 are all empty.
    # In golden_env, all 29 rows should contain VLOOKUP formulas.
    # -------------------------------------------------------------------------
    try:
        rows_with_vlookup = 0
        rows_with_any_formula = 0
        missing_rows = []

        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=4).value  # Column D
            if cell_val is None:
                missing_rows.append(row)
            elif isinstance(cell_val, str) and 'VLOOKUP' in cell_val.upper():
                rows_with_vlookup += 1
            elif isinstance(cell_val, str) and cell_val.startswith('='):
                rows_with_any_formula += 1
                missing_rows.append(row)
            else:
                missing_rows.append(row)

        expected_rows = len(DATA_ROWS)  # 29
        if rows_with_vlookup == expected_rows:
            print(f"PASS: Component 1 — All {expected_rows} rows in D2:D30 have VLOOKUP formulas (0.5 pts)")
            total_score += 0.5
        elif rows_with_vlookup > 0:
            partial = round(0.5 * rows_with_vlookup / expected_rows, 4)
            print(f"PARTIAL: Component 1 — {rows_with_vlookup}/{expected_rows} rows have VLOOKUP formulas ({partial} pts)")
            print(f"  Missing VLOOKUP in rows: {missing_rows[:10]}{'...' if len(missing_rows) > 10 else ''}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas found in column D rows 2-30")
            print(f"  Sample D2 value: {repr(ws.cell(row=2, column=4).value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: VLOOKUP formulas use approximate match (4th argument = 1 or TRUE)
    # (0.3 points)
    # Approximate match is required because the grade scale uses cutoff ranges.
    # The formula must have 4th argument as 1 or TRUE (not 0 or FALSE).
    # -------------------------------------------------------------------------
    try:
        rows_with_approx_match = 0
        rows_with_exact_match = 0
        rows_checked = 0

        # Pattern to find VLOOKUP with 4th argument
        # =VLOOKUP(lookup_value, table_array, col_index, [range_lookup])
        # range_lookup: 1 or TRUE = approximate match (sorted ascending)
        approx_pattern = re.compile(
            r'=VLOOKUP\s*\([^,]+,[^,]+,[^,]+,\s*(1|TRUE)\s*\)',
            re.IGNORECASE
        )
        exact_pattern = re.compile(
            r'=VLOOKUP\s*\([^,]+,[^,]+,[^,]+,\s*(0|FALSE)\s*\)',
            re.IGNORECASE
        )

        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str) and 'VLOOKUP' in cell_val.upper():
                rows_checked += 1
                if approx_pattern.search(cell_val):
                    rows_with_approx_match += 1
                elif exact_pattern.search(cell_val):
                    rows_with_exact_match += 1

        if rows_checked == 0:
            print(f"FAIL: Component 2 — No VLOOKUP formulas found to check for approximate match")
        elif rows_with_approx_match == rows_checked:
            print(f"PASS: Component 2 — All {rows_checked} VLOOKUP formulas use approximate match (4th arg=1/TRUE) (0.3 pts)")
            total_score += 0.3
        elif rows_with_approx_match > 0:
            partial = round(0.3 * rows_with_approx_match / rows_checked, 4)
            print(f"PARTIAL: Component 2 — {rows_with_approx_match}/{rows_checked} VLOOKUP formulas use approximate match ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — {rows_checked} VLOOKUP formula(s) found but none use approximate match")
            if rows_with_exact_match > 0:
                print(f"  {rows_with_exact_match} use exact match (0/FALSE) instead of approximate match")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Reference table F2:G6 is intact with correct grade scale
    # (0.2 points)
    # This checks that the reference table used by VLOOKUP contains the correct
    # score cutoffs and letter grades. This data exists in both initial and golden,
    # BUT we combine it as a sub-condition: VLOOKUP formulas reference $F$2:$G$6
    # AND the reference table data is correct. The formula reference check is
    # specific to the golden file since initial has no formulas.
    # -------------------------------------------------------------------------
    try:
        # Check that VLOOKUP formulas reference the grade scale table
        formulas_reference_table = 0
        rows_with_vlookup_for_check = 0

        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is not None and isinstance(cell_val, str) and 'VLOOKUP' in cell_val.upper():
                rows_with_vlookup_for_check += 1
                # Check if formula references F2:G6 (absolute or relative)
                if re.search(r'\$?F\$?2:\$?G\$?6', cell_val, re.IGNORECASE):
                    formulas_reference_table += 1

        # Check reference table data integrity
        grade_scale_issues = []
        expected_rows_data = [(2, 0, 'F'), (3, 60, 'D'), (4, 70, 'C'), (5, 80, 'B'), (6, 90, 'A')]
        for row, expected_score, expected_grade in expected_rows_data:
            f_val = ws.cell(row=row, column=6).value  # Column F
            g_val = ws.cell(row=row, column=7).value  # Column G
            if f_val != expected_score or g_val != expected_grade:
                grade_scale_issues.append(
                    f"Row {row}: F{row}={repr(f_val)} (expected {expected_score}), "
                    f"G{row}={repr(g_val)} (expected '{expected_grade}')"
                )

        grade_scale_error_count = len(grade_scale_issues)

        if rows_with_vlookup_for_check > 0 and formulas_reference_table == rows_with_vlookup_for_check and grade_scale_error_count == 0:
            print(f"PASS: Component 3 — All VLOOKUP formulas reference $F$2:$G$6 and grade scale is intact (0.2 pts)")
            total_score += 0.2
        elif rows_with_vlookup_for_check == 0:
            print(f"FAIL: Component 3 — No VLOOKUP formulas to check reference table usage")
        elif formulas_reference_table < rows_with_vlookup_for_check:
            print(f"FAIL: Component 3 — Only {formulas_reference_table}/{rows_with_vlookup_for_check} formulas reference F2:G6 table")
        elif grade_scale_error_count > 0:
            print(f"FAIL: Component 3 — Grade scale reference table has {grade_scale_error_count} issue(s):")
            for issue in grade_scale_issues:
                print(f"  {issue}")
        else:
            print(f"FAIL: Component 3 — Unexpected state: {formulas_reference_table} formulas reference table, {grade_scale_error_count} grade scale errors")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
