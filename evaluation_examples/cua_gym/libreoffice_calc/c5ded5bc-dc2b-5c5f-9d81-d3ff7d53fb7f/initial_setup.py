"""
Initial Setup: VLOOKUP grade lookup task — create student scores spreadsheet
Task ID: osworld_calc_vlookup_grade_lookup_001
Domain: libreoffice_calc

Creates a spreadsheet with:
  - Column A: StudentID, Column B: Name, Column C: Score, Column D: Grade (EMPTY)
  - Separate grade scale reference table in columns F-G
  - Grade column (D) MUST be empty — agent must add VLOOKUP formulas
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_grade_lookup_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Students ----
    ws = wb.active
    ws.title = "Students"

    # Headers in row 1
    headers = ["StudentID", "Name", "Score", "Grade"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 29 rows of realistic student data (rows 2-30)
    # Scores spread across all grade brackets
    students = [
        ("S001", "Emma Richardson",   92),
        ("S002", "Liam Nakamura",      78),
        ("S003", "Priya Patel",        85),
        ("S004", "Carlos Morales",     63),
        ("S005", "Sophia Okonkwo",     71),
        ("S006", "James Whitfield",    55),
        ("S007", "Aisha Diallo",       90),
        ("S008", "Noah Fernandez",     47),
        ("S009", "Mei-Ling Zhou",      83),
        ("S010", "Dylan O'Brien",      96),
        ("S011", "Fatima Al-Hassan",   68),
        ("S012", "Ryan Kowalski",      74),
        ("S013", "Zoe Bergmann",       39),
        ("S014", "Marcus Adeyemi",     81),
        ("S015", "Hannah Johansson",   58),
        ("S016", "Andre Baptiste",     93),
        ("S017", "Elena Vasquez",      77),
        ("S018", "Rohan Mehta",        62),
        ("S019", "Isabel Ferreira",    88),
        ("S020", "Tyler Nguyen",       44),
        ("S021", "Grace Amara",        70),
        ("S022", "Samuel Kim",         95),
        ("S023", "Nadia Petrova",      66),
        ("S024", "Owen Castillo",      82),
        ("S025", "Lily Tanaka",        51),
        ("S026", "Elijah Brooks",      79),
        ("S027", "Clara Huber",        87),
        ("S028", "Mateo Reyes",        34),
        ("S029", "Amara Nwosu",        91),
    ]

    for r, (sid, name, score) in enumerate(students, 2):
        ws.cell(row=r, column=1, value=sid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=score)
        # Column D (Grade) — intentionally left EMPTY

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10

    # ---- Grade Scale reference table in columns F-G ----
    ws.cell(row=1, column=6, value="Score").font = Font(bold=True)
    ws.cell(row=1, column=7, value="Grade").font = Font(bold=True)

    grade_scale = [
        (0,  "F"),
        (60, "D"),
        (70, "C"),
        (80, "B"),
        (90, "A"),
    ]
    for r, (score_cut, grade) in enumerate(grade_scale, 2):
        ws.cell(row=r, column=6, value=score_cut)
        ws.cell(row=r, column=7, value=grade)

    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
