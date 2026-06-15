"""
Reward Script: Enter IFERROR(VLOOKUP(...)) formula in D2 of Employee Lookup sheet
Task ID: calc_fmb_iferror_vlookup_046
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): D2 contains any formula (starts with '=')
  Component 2 (0.4): Formula uses both IFERROR and VLOOKUP with correct lookup range
  Component 3 (0.3): Formula exactly matches expected =IFERROR(VLOOKUP(A2,$F$2:$G$20,2,0),"Unknown")
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_iferror_vlookup_046'
SHEET_NAME = 'Employee Lookup'
TARGET_CELL = 'D2'
EXPECTED_FORMULA = '=IFERROR(VLOOKUP(A2,$F$2:$G$20,2,0),"Unknown")'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, no spaces."""
    if not f:
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Retrieve the value in D2
    d2_value = ws[TARGET_CELL].value

    # Component 1: D2 contains a formula (starts with '=') (0.3 points)
    try:
        has_formula = isinstance(d2_value, str) and d2_value.strip().startswith('=')
        if has_formula:
            print(f"PASS: Component 1 — D2 contains a formula: {repr(d2_value)} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — D2 does not contain a formula; found: {repr(d2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula contains IFERROR wrapping VLOOKUP with correct lookup range
    # and correct column index and exact-match flag (0.4 points)
    try:
        if isinstance(d2_value, str):
            norm = normalize_formula(d2_value)
            has_iferror = 'IFERROR(' in norm
            has_vlookup = 'VLOOKUP(' in norm
            has_range = '$F$2:$G$20' in d2_value.upper()
            has_col_index = ',2,0)' in norm or ',2,FALSE)' in norm or ',2,0,' in norm
            has_unknown = '"UNKNOWN"' in norm or "'UNKNOWN'" in norm

            if has_iferror and has_vlookup and has_range and has_col_index and has_unknown:
                print(f"PASS: Component 2 — Formula has IFERROR+VLOOKUP with correct range/col/flag/fallback (0.4 pts)")
                total_score += 0.4
            else:
                details = []
                if not has_iferror:
                    details.append("missing IFERROR")
                if not has_vlookup:
                    details.append("missing VLOOKUP")
                if not has_range:
                    details.append("wrong or missing lookup range (expected $F$2:$G$20)")
                if not has_col_index:
                    details.append("wrong column index or match type (expected 2,0)")
                if not has_unknown:
                    details.append('missing "Unknown" error value')
                print(f"FAIL: Component 2 — {'; '.join(details)}; formula: {repr(d2_value)}")
        else:
            print(f"FAIL: Component 2 — D2 value is not a string formula: {repr(d2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula exactly matches the expected formula (case-insensitive) (0.3 points)
    try:
        if isinstance(d2_value, str):
            norm_actual = normalize_formula(d2_value)
            norm_expected = normalize_formula(EXPECTED_FORMULA)
            if norm_actual == norm_expected:
                print(f"PASS: Component 3 — Formula exactly matches expected (case-insensitive) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — Formula mismatch")
                print(f"  Expected (normalized): {norm_expected}")
                print(f"  Actual   (normalized): {norm_actual}")
        else:
            print(f"FAIL: Component 3 — D2 value is not a formula string: {repr(d2_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Also verify no other cells were modified (informational, no score deduction)
    try:
        # Check D3 is still empty (it should be None in both initial and golden)
        d3_value = ws['D3'].value
        if d3_value is not None:
            print(f"INFO: D3 is not empty (value: {repr(d3_value)}) — task specified only D2 should change")
    except Exception as e:
        print(f"INFO: Could not check D3: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
