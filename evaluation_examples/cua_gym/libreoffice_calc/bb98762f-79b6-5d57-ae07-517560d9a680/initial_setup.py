"""
Initial Setup: Employee lookup spreadsheet with IFERROR/VLOOKUP task
Task ID: calc_fmb_iferror_vlookup_046
Domain: libreoffice_calc

Creates an employee spreadsheet with a main table (A1:D31) and a lookup table
(F1:G20). A2 contains an employee ID that does NOT exist in the lookup table.
D2 is intentionally left empty (target cell for the task).
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_iferror_vlookup_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Employee Lookup'

    # --- Main table headers ---
    ws['A1'] = 'Emp ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Salary'
    ws['D1'] = 'Department'

    # --- Lookup table headers ---
    ws['F1'] = 'Lookup ID'
    ws['G1'] = 'Department'

    # Departments available in the lookup table
    departments = [
        'Engineering',
        'Marketing',
        'Sales',
        'Finance',
        'Human Resources',
        'Operations',
        'Legal',
        'Product',
        'Customer Support',
        'Research',
        'Design',
        'Procurement',
        'IT',
        'Compliance',
        'Public Relations',
        'Data Science',
        'Business Development',
        'Quality Assurance',
    ]

    # Lookup IDs that ARE in the table: E101-E118 (18 entries, rows 2-19)
    # We leave row 20 for one more lookup entry: E120
    lookup_ids = [
        'E101', 'E102', 'E103', 'E104', 'E105', 'E106',
        'E107', 'E108', 'E109', 'E110', 'E111', 'E112',
        'E113', 'E114', 'E115', 'E116', 'E117', 'E118',
        'E120',
    ]
    for i, (lid, dept) in enumerate(zip(lookup_ids, departments), start=2):
        ws.cell(row=i, column=6, value=lid)
        ws.cell(row=i, column=7, value=dept)

    # --- Employee records ---
    # A2 gets a special ID that is NOT in the lookup table (E999)
    # Other rows get IDs from lookup_ids that DO exist
    employee_data = [
        # (Emp ID, Name, Salary) — D column left empty
        ('E999',  'Jordan Blake',       52000),   # row 2 — NOT in lookup
        ('E101',  'Sarah Chen',         85000),
        ('E102',  'Marcus Johnson',     72000),
        ('E103',  'Priya Patel',        91000),
        ('E104',  'Liam O\'Brien',      67000),
        ('E105',  'Nadia Kowalski',     79500),
        ('E106',  'James Thornton',     88000),
        ('E107',  'Aisha Mohammed',     62000),
        ('E108',  'Carlos Rivera',      74000),
        ('E109',  'Mei-Lin Zhang',      95000),
        ('E110',  'David Okonkwo',      58000),
        ('E111',  'Fatima Al-Hassan',   83000),
        ('E112',  'Oliver Schmidt',     76500),
        ('E113',  'Yuki Tanaka',        69000),
        ('E114',  'Elena Petrova',      81000),
        ('E115',  'Samuel Oduya',       55000),
        ('E116',  'Ingrid Lindqvist',   92000),
        ('E117',  'Rafael Mendes',      70000),
        ('E118',  'Chloe Dubois',       78000),
        ('E120',  'Ahmed Khalil',       64000),
        ('E101',  'Sophia Turner',      87500),
        ('E103',  'Ben Nakamura',       93000),
        ('E105',  'Aaliya Singh',       80000),
        ('E107',  'Derek Williams',     61000),
        ('E109',  'Leona Vasquez',      96000),
        ('E111',  'Patrick O\'Connor',  84000),
        ('E113',  'Vivienne Laurent',   71000),
        ('E115',  'Kofi Asante',        57000),
        ('E117',  'Miriam Goldstein',   68000),
        ('E120',  'Thomas Bergmann',    65000),
    ]

    for row_idx, (emp_id, name, salary) in enumerate(employee_data, start=2):
        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=salary)
        # Column D (Department) intentionally left empty for ALL rows
        # to represent pre-filled data; D2 is the task target and must stay empty.

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
