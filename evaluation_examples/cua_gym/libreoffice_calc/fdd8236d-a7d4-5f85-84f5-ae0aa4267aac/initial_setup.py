"""
Initial Setup: Help desk ticket analytics report
Task ID: calc_wf_081
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Tickets ---
    ws = wb.active
    ws.title = 'Tickets'

    headers = [
        'Ticket ID', 'Created Date', 'Category', 'Priority',
        'Assigned To', 'Resolved Date', 'Resolution Hours',
        'SLA Target Hours', 'CSAT'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align

    categories = ['Software', 'Hardware', 'Network', 'Access']
    sla_targets = {'P1': 4, 'P2': 8, 'P3': 24, 'P4': 48}
    priorities = ['P1', 'P2', 'P3', 'P4']
    priority_weights = [0.15, 0.25, 0.35, 0.25]

    agents = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'David Kim',
        'Emma Rodriguez', 'James Wilson', 'Aisha Mohammed', 'Carlos Silva',
        'Lisa Thompson', 'Ryan O\'Brien'
    ]

    # Generate 75 tickets spread over ~60 days
    base_date = datetime(2025, 1, 6)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    for i in range(75):
        row = i + 2
        ticket_id = f'TKT-{1001 + i}'
        created = base_date + timedelta(days=random.randint(0, 59),
                                         hours=random.randint(8, 17),
                                         minutes=random.randint(0, 59))

        priority = random.choices(priorities, weights=priority_weights, k=1)[0]
        category = random.choice(categories)
        agent = random.choice(agents)
        sla_target = sla_targets[priority]

        # Resolution time: mostly within SLA, some breaches (~25%)
        if random.random() < 0.75:
            resolution_hours = round(random.uniform(0.5, sla_target * 0.95), 1)
        else:
            resolution_hours = round(random.uniform(sla_target * 1.1, sla_target * 3.0), 1)

        resolved = created + timedelta(hours=resolution_hours)
        csat = random.choices([1, 2, 3, 4, 5], weights=[0.05, 0.10, 0.20, 0.35, 0.30], k=1)[0]

        values = [
            ticket_id, created, category, priority,
            agent, resolved, resolution_hours,
            sla_target, csat
        ]
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            c.border = thin_border
            if col in (2, 6):  # date columns
                c.number_format = 'yyyy-mm-dd hh:mm'
            elif col == 7:
                c.number_format = '0.0'
            elif col == 8:
                c.number_format = '0'
            elif col == 9:
                c.number_format = '0'

    # Column widths
    col_widths = {'A': 12, 'B': 18, 'C': 12, 'D': 10, 'E': 18,
                  'F': 18, 'G': 16, 'H': 16, 'I': 8}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
