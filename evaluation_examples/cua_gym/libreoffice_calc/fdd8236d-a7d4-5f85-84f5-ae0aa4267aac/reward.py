"""
Reward Script: Help desk ticket analytics report
Task ID: calc_wf_081
Domain: libreoffice_calc
Scoring:
  Component 1: SLA Met column (J) with IF formulas — 0.20 pts
  Component 2: Conditional formatting on SLA breaches — 0.15 pts
  Component 3: Summary sheet with SLA compliance stats — 0.15 pts
  Component 4: Avg resolution by category (AVERAGEIFS) — 0.15 pts
  Component 5: Avg resolution by priority (AVERAGEIFS) — 0.10 pts
  Component 6: Pivot-style COUNTIFS table — 0.15 pts
  Component 7: Charts (line + bar) — 0.10 pts
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_081'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Tickets sheet must exist with original data
    if 'Tickets' not in wb.sheetnames:
        print("CRITICAL: 'Tickets' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_tickets = wb['Tickets']

    # Component 1: SLA Met column (J) with IF formulas (0.20 points)
    # Task requires: SLA Met = IF(Resolution <= Target, "Yes", "No") in column J
    # This column does NOT exist in initial_env (only 9 columns A-I)
    try:
        j_header = ws_tickets.cell(row=1, column=10).value
        has_sla_header = (j_header is not None and
                          'sla' in str(j_header).lower() and
                          'met' in str(j_header).lower())

        if has_sla_header:
            # Check that at least 50 cells in J have IF formulas comparing G and H
            formula_count = 0
            for r in range(2, 77):  # rows 2-76 (75 tickets)
                val = ws_tickets.cell(row=r, column=10).value
                if val is not None and isinstance(val, str):
                    val_upper = val.upper().replace(" ", "")
                    if '=IF(' in val_upper and 'G' in val_upper and 'H' in val_upper:
                        formula_count += 1
            if formula_count >= 50:
                print(f"PASS: Component 1 — SLA Met column with {formula_count} IF formulas (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — Only {formula_count} IF formulas in column J (need >= 50)")
        else:
            print(f"FAIL: Component 1 — Column J header is '{j_header}', expected 'SLA Met'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Conditional formatting on SLA breaches (0.15 points)
    # Task requires: conditional formatting to highlight when Resolution Hours > SLA Target Hours
    # Initial file has NO conditional formatting
    try:
        cf_rules = list(ws_tickets.conditional_formatting)
        found_sla_cf = False
        for cf in cf_rules:
            cf_range = str(cf)
            for rule in cf.rules:
                # Looking for a rule that compares G (resolution) vs H (target)
                # or highlights when resolution exceeds target
                rule_formula = getattr(rule, 'formula', None)
                rule_operator = getattr(rule, 'operator', None)
                rule_type = getattr(rule, 'type', None)

                # Check if the CF applies to column G (Resolution Hours) area
                if 'G' in cf_range or 'J' in cf_range or ':' in cf_range:
                    # Check for red-ish fill in the differential style
                    has_red_fill = False
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fg_rgb = rule.dxf.fill.fgColor.rgb
                            if fg_rgb and 'FF' in fg_rgb[:2] and 'FF' in fg_rgb[2:4]:
                                has_red_fill = True
                            # Also accept other red-ish variants
                            if fg_rgb and fg_rgb.upper().startswith('FFFF0'):
                                has_red_fill = True
                        except Exception:
                            pass

                    # Accept: cellIs with greaterThan on H column, or formula-based
                    if rule_type == 'cellIs' and rule_operator == 'greaterThan':
                        found_sla_cf = True
                    elif rule_type == 'expression' or rule_type == 'formula':
                        found_sla_cf = True
                    elif has_red_fill:
                        found_sla_cf = True

        if found_sla_cf:
            print(f"PASS: Component 2 — Conditional formatting for SLA breaches found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — No conditional formatting for SLA breaches found (rules: {len(cf_rules)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary sheet with SLA compliance stats (0.15 points)
    # Initial file has NO Summary sheet
    try:
        has_summary = False
        summary_ws = None
        for sn in wb.sheetnames:
            if 'summary' in sn.lower():
                has_summary = True
                summary_ws = wb[sn]
                break

        if has_summary and summary_ws is not None:
            # Look for SLA compliance rate formula (COUNTIF-based)
            found_sla_rate = False
            for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row,
                                             min_col=1, max_col=summary_ws.max_column):
                for cell in row:
                    val = cell.value
                    if val and isinstance(val, str) and 'COUNTIF' in val.upper():
                        # Found a COUNTIF formula referencing SLA data
                        found_sla_rate = True
                        break
                if found_sla_rate:
                    break

            if found_sla_rate:
                print(f"PASS: Component 3 — Summary sheet with SLA compliance stats (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — Summary sheet exists but no COUNTIF formula for SLA compliance")
        else:
            print(f"FAIL: Component 3 — No Summary sheet found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Average resolution by category using AVERAGEIFS (0.15 points)
    # Task requires AVERAGEIFS for Software, Hardware, Network, Access
    # Initial file has NO such formulas
    try:
        if summary_ws is not None:
            categories = ['software', 'hardware', 'network', 'access']
            found_categories = set()
            for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row,
                                             min_col=1, max_col=summary_ws.max_column):
                for cell in row:
                    val = cell.value
                    if val and isinstance(val, str) and 'AVERAGEIF' in val.upper():
                        val_lower = val.lower()
                        for cat in categories:
                            if cat in val_lower:
                                found_categories.add(cat)

            if len(found_categories) >= 3:
                print(f"PASS: Component 4 — AVERAGEIFS for categories: {found_categories} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 — Only found AVERAGEIFS for {found_categories}, need >= 3 of {categories}")
        else:
            print(f"FAIL: Component 4 — No Summary sheet to check AVERAGEIFS")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Average resolution by priority using AVERAGEIFS (0.10 points)
    # Task requires AVERAGEIFS for P1, P2, P3, P4
    try:
        if summary_ws is not None:
            priorities = ['p1', 'p2', 'p3', 'p4']
            found_priorities = set()
            for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row,
                                             min_col=1, max_col=summary_ws.max_column):
                for cell in row:
                    val = cell.value
                    if val and isinstance(val, str) and 'AVERAGEIF' in val.upper():
                        val_lower = val.lower()
                        for pri in priorities:
                            if f'"{pri}"' in val_lower or f"'{pri}'" in val_lower:
                                found_priorities.add(pri)

            if len(found_priorities) >= 3:
                print(f"PASS: Component 5 — AVERAGEIFS for priorities: {found_priorities} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 — Only found AVERAGEIFS for {found_priorities}, need >= 3 of {priorities}")
        else:
            print(f"FAIL: Component 5 — No Summary sheet to check AVERAGEIFS by priority")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Pivot-style summary table using COUNTIFS (0.15 points)
    # Task requires: categories as rows, priorities as columns, counts as values
    # Uses COUNTIFS formulas
    try:
        if summary_ws is not None:
            countifs_count = 0
            for row in summary_ws.iter_rows(min_row=1, max_row=summary_ws.max_row,
                                             min_col=1, max_col=summary_ws.max_column):
                for cell in row:
                    val = cell.value
                    if val and isinstance(val, str) and 'COUNTIFS(' in val.upper():
                        countifs_count += 1

            # Expect at least 12 COUNTIFS cells (4 categories x 3+ priorities min)
            if countifs_count >= 8:
                print(f"PASS: Component 6 — Pivot-style table with {countifs_count} COUNTIFS formulas (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 6 — Only {countifs_count} COUNTIFS formulas, need >= 8 for pivot table")
        else:
            print(f"FAIL: Component 6 — No Summary sheet for pivot table")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Charts — line chart (ticket volume) + bar chart (avg resolution by category) (0.10 points)
    # Initial file has NO charts
    try:
        # Charts can be on any sheet
        total_charts = 0
        has_line_chart = False
        has_bar_chart = False
        for sn in wb.sheetnames:
            ws = wb[sn]
            for chart in ws._charts:
                total_charts += 1
                chart_type = type(chart).__name__
                if 'Line' in chart_type:
                    has_line_chart = True
                elif 'Bar' in chart_type:
                    has_bar_chart = True

        if has_line_chart and has_bar_chart:
            print(f"PASS: Component 7 — Found line chart and bar chart ({total_charts} total) (0.10 pts)")
            total_score += 0.10
        elif total_charts >= 2:
            # Accept 2+ charts even if types don't exactly match
            print(f"PASS: Component 7 — Found {total_charts} charts (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 — Found {total_charts} chart(s), line={has_line_chart}, bar={has_bar_chart}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
