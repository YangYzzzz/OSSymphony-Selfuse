"""
Initial Setup: Create customer database spreadsheet with no freeze panes
Task ID: calc_ggf_004
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Database'

    # --- Headers (Row 1, Columns A-R) ---
    headers = [
        'Customer ID', 'Full Name', 'Email', 'Phone', 'City', 'State',
        'Country', 'Tier', 'Joined Date', 'Last Purchase', 'Total Spent',
        'Orders', 'Rating', 'Preferred Contact', 'Newsletter', 'Status',
        'Region', 'Notes'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data generation pools ---
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
        'Carlos', 'Fatima', 'Robert', 'Aisha', 'Thomas', 'Yuki', 'Michael',
        'Zara', 'William', 'Sofia', 'Daniel', 'Amara', 'Christopher',
        'Olivia', 'Benjamin', 'Emma', 'Alexander', 'Chloe', 'Ethan',
        'Isabella', 'Nathan', 'Mia', 'Lucas', 'Aria', 'Ryan', 'Lily',
        'Andrew', 'Nora', 'Kevin', 'Hannah', 'Patrick', 'Grace', 'Derek',
        'Victoria', 'Brandon', 'Samantha', 'Tyler', 'Rachel', 'Jason',
        'Lauren', 'Keith', 'Monica', 'Trevor',
    ]
    last_names = [
        'Chen', 'Johnson', 'Patel', 'Williams', 'Singh', 'Kim', 'Brown',
        'Garcia', 'Ali', 'Smith', 'Nguyen', 'Martinez', 'Tanaka', 'Davis',
        'Wilson', 'Anderson', 'Taylor', 'Thomas', 'Lee', 'White',
        'Harris', 'Clark', 'Lewis', 'Young', 'King', 'Wright', 'Hill',
        'Scott', 'Green', 'Adams', 'Baker', 'Nelson', 'Carter', 'Mitchell',
        'Roberts', 'Turner', 'Phillips', 'Campbell', 'Parker', 'Evans',
        'Edwards', 'Collins', 'Stewart', 'Morris', 'Murphy', 'Rivera',
        'Cook', 'Rogers', 'Morgan', 'Cooper',
    ]
    cities = [
        ('New York', 'NY'), ('Los Angeles', 'CA'), ('Chicago', 'IL'),
        ('Houston', 'TX'), ('Phoenix', 'AZ'), ('Philadelphia', 'PA'),
        ('San Antonio', 'TX'), ('San Diego', 'CA'), ('Dallas', 'TX'),
        ('San Jose', 'CA'), ('Austin', 'TX'), ('Jacksonville', 'FL'),
        ('Fort Worth', 'TX'), ('Columbus', 'OH'), ('Charlotte', 'NC'),
        ('Indianapolis', 'IN'), ('San Francisco', 'CA'), ('Seattle', 'WA'),
        ('Denver', 'CO'), ('Nashville', 'TN'), ('Portland', 'OR'),
        ('Miami', 'FL'), ('Atlanta', 'GA'), ('Boston', 'MA'),
        ('Minneapolis', 'MN'), ('Detroit', 'MI'), ('Tampa', 'FL'),
        ('Pittsburgh', 'PA'), ('St. Louis', 'MO'), ('Baltimore', 'MD'),
    ]
    countries = ['United States']
    tiers = ['Bronze', 'Silver', 'Gold', 'Platinum', 'Diamond']
    tier_weights = [35, 30, 20, 10, 5]
    contact_methods = ['Email', 'Phone', 'SMS', 'Mail']
    statuses = ['Active', 'Inactive', 'Suspended', 'Pending']
    status_weights = [70, 15, 5, 10]
    note_templates = [
        'Preferred customer since {year}',
        'Referred by existing member',
        'Upgraded from {old_tier} tier',
        'Frequent buyer - {cat} category',
        'VIP event attendee',
        'Corporate account holder',
        'Seasonal purchaser',
        'High-value returns noted',
        'Loyalty program participant',
        '',  # some rows have no notes
        '',
        '',
    ]
    categories = ['Electronics', 'Apparel', 'Home & Garden', 'Sports', 'Books', 'Beauty']

    # --- Generate 600 rows of customer data ---
    for i in range(600):
        row = i + 2  # data starts at row 2
        cust_id = f'CID-{10001 + i}'
        first = random.choice(first_names)
        last = random.choice(last_names)
        full_name = f'{first} {last}'
        email = f'{first.lower()}.{last.lower()}{random.randint(1, 99)}@{"gmail.com" if random.random() < 0.4 else "outlook.com" if random.random() < 0.5 else "yahoo.com"}'
        phone = f'({random.randint(200,999)}) {random.randint(100,999)}-{random.randint(1000,9999)}'
        city, state = random.choice(cities)
        country = 'United States'
        tier = random.choices(tiers, weights=tier_weights, k=1)[0]

        # Joined date between 2018-01-01 and 2025-12-31
        join_year = random.randint(2018, 2025)
        join_month = random.randint(1, 12)
        join_day = random.randint(1, 28)
        joined_date = f'{join_year}-{join_month:02d}-{join_day:02d}'

        # Last purchase date (after join date, up to 2026-03-15)
        lp_year = random.randint(max(join_year, 2024), 2026)
        lp_month = random.randint(1, 12) if lp_year < 2026 else random.randint(1, 3)
        lp_day = random.randint(1, 28)
        last_purchase = f'{lp_year}-{lp_month:02d}-{lp_day:02d}'

        # Spending based on tier
        tier_spend_range = {
            'Bronze': (150, 5000),
            'Silver': (3000, 15000),
            'Gold': (10000, 50000),
            'Platinum': (30000, 120000),
            'Diamond': (80000, 500000),
        }
        lo, hi = tier_spend_range[tier]
        total_spent = round(random.uniform(lo, hi), 2)

        orders = random.randint(1, 80) if tier != 'Diamond' else random.randint(20, 200)
        rating = round(random.uniform(1.0, 5.0), 1)
        preferred_contact = random.choice(contact_methods)
        newsletter = random.choice(['Yes', 'No'])
        status = random.choices(statuses, weights=status_weights, k=1)[0]

        # Region based on state
        east_states = ['NY', 'PA', 'FL', 'NC', 'GA', 'MA', 'MD', 'IN', 'OH']
        west_states = ['CA', 'AZ', 'WA', 'OR', 'CO']
        central_states = ['IL', 'TX', 'MN', 'MI', 'MO', 'TN']
        if state in east_states:
            region = 'East'
        elif state in west_states:
            region = 'West'
        else:
            region = 'Central'

        note_tmpl = random.choice(note_templates)
        if '{year}' in note_tmpl:
            notes = note_tmpl.format(year=join_year)
        elif '{old_tier}' in note_tmpl:
            idx = tiers.index(tier)
            old = tiers[max(0, idx - 1)]
            notes = note_tmpl.format(old_tier=old)
        elif '{cat}' in note_tmpl:
            notes = note_tmpl.format(cat=random.choice(categories))
        else:
            notes = note_tmpl

        row_data = [
            cust_id, full_name, email, phone, city, state, country,
            tier, joined_date, last_purchase, total_spent, orders,
            rating, preferred_contact, newsletter, status, region, notes
        ]
        for c, val in enumerate(row_data, 1):
            ws.cell(row=row, column=c, value=val)

    # --- Column widths for readability ---
    col_widths = {
        'A': 14, 'B': 22, 'C': 32, 'D': 18, 'E': 16, 'F': 8,
        'G': 16, 'H': 10, 'I': 14, 'J': 14, 'K': 14, 'L': 8,
        'M': 8, 'N': 18, 'O': 12, 'P': 10, 'Q': 10, 'R': 34,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row 1 height
    ws.row_dimensions[1].height = 20

    # NO freeze panes - this is the task the agent must complete
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
