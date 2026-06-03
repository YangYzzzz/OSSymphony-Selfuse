"""
Reward Script: Freeze first row and first two columns on 'Database' sheet
Task ID: calc_ggf_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): freeze_panes is set (not None) on Database sheet
  Component 2 (0.5): freeze_panes is exactly 'C2' (row 1 + cols A-B frozen)
"""

import os

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_004'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify freeze panes on the Database sheet.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Database' sheet must exist
    if 'Database' not in wb.sheetnames:
        print(f"CRITICAL: 'Database' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Database']

    # Component 1: freeze_panes is set (not None) on Database sheet (0.5 points)
    try:
        fp = ws.freeze_panes
        if fp is not None:
            print(f"PASS: Component 1 — freeze_panes is set to '{fp}' (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — freeze_panes is None (no freeze applied)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: freeze_panes is exactly 'C2' (freezes row 1 + columns A,B) (0.5 points)
    try:
        fp = ws.freeze_panes
        if fp is not None and str(fp) == 'C2':
            print(f"PASS: Component 2 — freeze_panes is exactly 'C2' (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — expected freeze_panes='C2', found '{fp}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
