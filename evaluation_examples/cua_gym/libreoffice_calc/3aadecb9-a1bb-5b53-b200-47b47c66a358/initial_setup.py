"""
Initial Setup: Purchase Order Tracking System
Task ID: calc_wf_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: Suppliers ==========
    ws_sup = wb.active
    ws_sup.title = 'Suppliers'

    sup_headers = ['Supplier ID', 'Supplier Name', 'Contact', 'Terms', 'Rating']
    for col, h in enumerate(sup_headers, 1):
        cell = ws_sup.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    suppliers = [
        ['S001', 'Apex Industrial Supply', 'karen.mitchell@apex-ind.com', 'Net 30', 4.5],
        ['S002', 'BrightWave Electronics', 'david.chen@brightwave.com', 'Net 60', 4.2],
        ['S003', 'CoreTech Components', 'lisa.park@coretech.com', 'Net 30', 4.8],
        ['S004', 'Delta Manufacturing Co.', 'robert.james@deltamfg.com', 'Net 90', 3.9],
        ['S005', 'Evergreen Materials', 'sarah.nguyen@evergreen-mat.com', 'Net 30', 4.6],
        ['S006', 'FusionParts Ltd.', 'michael.brown@fusionparts.com', 'Net 60', 4.1],
        ['S007', 'GlobalSource Trading', 'jennifer.wu@globalsource.com', 'Net 30', 4.7],
        ['S008', 'HarborLine Logistics', 'thomas.garcia@harborline.com', 'Net 90', 3.8],
        ['S009', 'InnovateTech Solutions', 'amanda.taylor@innovatetech.com', 'Net 60', 4.4],
        ['S010', 'JetStream Supplies', 'kevin.lee@jetstream-sup.com', 'Net 30', 4.3],
        ['S011', 'KingPin Fasteners', 'rachel.adams@kingpin-fast.com', 'Net 60', 4.0],
        ['S012', 'LuminaPower Systems', 'daniel.martinez@luminapower.com', 'Net 90', 4.5],
    ]

    for r, row_data in enumerate(suppliers, 2):
        for c, val in enumerate(row_data, 1):
            ws_sup.cell(row=r, column=c, value=val)

    # Column widths for Suppliers
    ws_sup.column_dimensions['A'].width = 14
    ws_sup.column_dimensions['B'].width = 28
    ws_sup.column_dimensions['C'].width = 36
    ws_sup.column_dimensions['D'].width = 12
    ws_sup.column_dimensions['E'].width = 10

    # ========== Sheet 2: PO Log ==========
    ws_po = wb.create_sheet('PO Log')

    po_headers = ['PO #', 'PO Date', 'Supplier ID', 'Supplier Name', 'Item',
                  'Qty', 'Unit Price', 'Discount Tier', 'Total',
                  'Expected Delivery', 'Actual Delivery', 'Status']
    for col, h in enumerate(po_headers, 1):
        cell = ws_po.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 25 PO entries - raw data only; formulas will be added by golden_patch
    # Columns: PO#, PO Date, Supplier ID, (Supplier Name-blank), Item, Qty, Unit Price,
    #          (Discount Tier-blank), (Total-blank), (Expected Delivery-blank),
    #          Actual Delivery (some filled), (Status-blank)
    base_date = date(2025, 9, 1)

    po_data = [
        ['PO-1001', date(2025, 9, 2),  'S001', 'Steel mounting brackets', 150, 12.50, None],
        ['PO-1002', date(2025, 9, 5),  'S003', 'Circuit board assemblies', 75, 45.00, None],
        ['PO-1003', date(2025, 9, 8),  'S002', 'LED display panels', 520, 28.75, None],
        ['PO-1004', date(2025, 9, 10), 'S005', 'Aluminum extrusion rods', 1200, 8.40, None],
        ['PO-1005', date(2025, 9, 12), 'S007', 'Copper wiring harness', 200, 15.60, None],
        ['PO-1006', date(2025, 9, 15), 'S004', 'Hydraulic actuators', 30, 285.00, None],
        ['PO-1007', date(2025, 9, 18), 'S006', 'Thermal paste cartridges', 600, 6.25, None],
        ['PO-1008', date(2025, 9, 20), 'S009', 'Fiber optic cables (50m)', 90, 52.00, None],
        ['PO-1009', date(2025, 9, 22), 'S001', 'Stainless steel bolts M8', 2500, 0.85, None],
        ['PO-1010', date(2025, 9, 25), 'S010', 'Rubber gasket seals', 350, 3.20, None],
        ['PO-1011', date(2025, 9, 28), 'S011', 'Hex cap screws assorted', 1500, 0.45, None],
        ['PO-1012', date(2025, 10, 1), 'S003', 'Microcontroller units', 180, 22.50, None],
        ['PO-1013', date(2025, 10, 3), 'S008', 'Shipping pallets (wooden)', 60, 18.00, None],
        ['PO-1014', date(2025, 10, 5), 'S012', 'Solar panel cells 5W', 800, 14.80, None],
        ['PO-1015', date(2025, 10, 8), 'S002', 'Capacitor banks 100uF', 1100, 2.10, None],
        ['PO-1016', date(2025, 10, 10), 'S005', 'Carbon fiber sheets', 45, 120.00, None],
        ['PO-1017', date(2025, 10, 12), 'S007', 'Insulated wire spools', 250, 9.75, None],
        ['PO-1018', date(2025, 10, 15), 'S006', 'Heat sink assemblies', 500, 7.50, None],
        ['PO-1019', date(2025, 10, 18), 'S009', 'USB-C connector ports', 3000, 1.20, None],
        ['PO-1020', date(2025, 10, 20), 'S004', 'Pneumatic cylinders', 25, 340.00, None],
        ['PO-1021', date(2025, 10, 22), 'S010', 'Silicone tubing (1m)', 700, 4.50, None],
        ['PO-1022', date(2025, 10, 25), 'S001', 'Welding electrodes pack', 400, 5.80, None],
        ['PO-1023', date(2025, 10, 28), 'S012', 'Battery cells 18650', 2000, 3.25, None],
        ['PO-1024', date(2025, 11, 1), 'S003', 'OLED screen modules', 60, 85.00, None],
        ['PO-1025', date(2025, 11, 3), 'S011', 'Spring washers M6', 5000, 0.15, None],
    ]

    # Actual delivery dates - some delivered, some not
    actual_deliveries = {
        0: date(2025, 10, 1),   # PO-1001 delivered
        1: date(2025, 10, 10),  # PO-1002 delivered
        2: date(2025, 11, 8),   # PO-1003 delivered
        3: date(2025, 10, 5),   # PO-1004 delivered
        4: date(2025, 10, 14),  # PO-1005 delivered
        6: date(2025, 11, 18),  # PO-1007 delivered
        7: date(2025, 11, 20),  # PO-1008 delivered
        8: date(2025, 10, 20),  # PO-1009 delivered
        9: date(2025, 10, 28),  # PO-1010 delivered
        10: date(2025, 11, 30), # PO-1011 delivered
        11: date(2025, 11, 2),  # PO-1012 delivered
        14: date(2025, 12, 15), # PO-1015 delivered
        # Remaining POs: no actual delivery yet
    }

    for i, po in enumerate(po_data):
        row = i + 2
        ws_po.cell(row=row, column=1, value=po[0])  # PO #
        ws_po.cell(row=row, column=2, value=po[1])   # PO Date
        ws_po.cell(row=row, column=2).number_format = 'yyyy-mm-dd'
        ws_po.cell(row=row, column=3, value=po[2])   # Supplier ID
        # Column 4 (Supplier Name) - intentionally blank, to be filled by VLOOKUP
        ws_po.cell(row=row, column=5, value=po[3])   # Item
        ws_po.cell(row=row, column=6, value=po[4])   # Qty
        ws_po.cell(row=row, column=7, value=po[5])   # Unit Price
        ws_po.cell(row=row, column=7).number_format = '$#,##0.00'
        # Columns 8 (Discount Tier), 9 (Total), 10 (Expected Delivery), 12 (Status) - blank
        # Column 11 (Actual Delivery)
        if i in actual_deliveries:
            ws_po.cell(row=row, column=11, value=actual_deliveries[i])
            ws_po.cell(row=row, column=11).number_format = 'yyyy-mm-dd'

    # Column widths for PO Log
    ws_po.column_dimensions['A'].width = 12
    ws_po.column_dimensions['B'].width = 14
    ws_po.column_dimensions['C'].width = 14
    ws_po.column_dimensions['D'].width = 28
    ws_po.column_dimensions['E'].width = 30
    ws_po.column_dimensions['F'].width = 10
    ws_po.column_dimensions['G'].width = 14
    ws_po.column_dimensions['H'].width = 14
    ws_po.column_dimensions['I'].width = 16
    ws_po.column_dimensions['J'].width = 18
    ws_po.column_dimensions['K'].width = 18
    ws_po.column_dimensions['L'].width = 12

    # ========== Sheet 3: Status Board ==========
    ws_status = wb.create_sheet('Status Board')

    ws_status.cell(row=1, column=1, value='Purchase Order Status Summary')
    ws_status.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_status.merge_cells('A1:C1')

    status_headers = ['Status', 'Count', 'Percentage']
    for col, h in enumerate(status_headers, 1):
        cell = ws_status.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Status categories - values blank, to be filled by formulas
    statuses = ['Delivered', 'Pending', 'Late']
    for i, s in enumerate(statuses):
        ws_status.cell(row=4 + i, column=1, value=s)

    ws_status.cell(row=7, column=1, value='Total POs')
    ws_status.cell(row=7, column=1).font = Font(bold=True)

    ws_status.column_dimensions['A'].width = 18
    ws_status.column_dimensions['B'].width = 12
    ws_status.column_dimensions['C'].width = 14

    # Save
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
