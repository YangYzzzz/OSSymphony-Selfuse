"""
Reward Script: Purchase Order Tracking System
Task ID: calc_wf_040
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP supplier names in PO Log col D (0.20)
  Component 2: Discount tier formulas in col H (0.15)
  Component 3: Total calculation formulas in col I (0.15)
  Component 4: Expected delivery formulas in col J (0.15)
  Component 5: Status formulas in col L (0.15)
  Component 6: Status Board summary formulas (0.10)
  Component 7: Conditional formatting for Late/Delivered (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_040'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    required_sheets = ['Suppliers', 'PO Log', 'Status Board']
    for sn in required_sheets:
        if sn not in wb.sheetnames:
            print(f"CRITICAL: Missing sheet '{sn}'")
            print("REWARD: 0.0")
            return 0.0

    ws_suppliers = wb['Suppliers']
    ws_po = wb['PO Log']
    ws_status = wb['Status Board']

    # Build supplier lookup from Suppliers sheet for cross-reference
    supplier_map = {}
    for r in range(2, ws_suppliers.max_row + 1):
        sid = ws_suppliers.cell(row=r, column=1).value
        name = ws_suppliers.cell(row=r, column=2).value
        if sid:
            supplier_map[str(sid)] = str(name) if name else ''

    PO_ROWS = range(2, 27)  # rows 2-26 = 25 POs

    # =========================================================================
    # Component 1: VLOOKUP supplier names in col D (0.20 points)
    # Initial: col D is empty. Golden: VLOOKUP formulas filling supplier names.
    # =========================================================================
    try:
        vlookup_count = 0
        for r in PO_ROWS:
            val = ws_po.cell(row=r, column=4).value  # col D = Supplier Name
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                if "VLOOKUP" in val_str:
                    vlookup_count += 1
                else:
                    # Could be a resolved value -- check if it matches expected supplier
                    supplier_id = str(ws_po.cell(row=r, column=3).value)
                    expected_name = supplier_map.get(supplier_id, '')
                    if expected_name and str(val).strip() == expected_name.strip():
                        vlookup_count += 1

        ratio = vlookup_count / 25.0
        if ratio >= 0.9:
            print(f"PASS: Component 1 — VLOOKUP supplier names: {vlookup_count}/25 filled (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = 0.20 * ratio
            print(f"PARTIAL: Component 1 — VLOOKUP supplier names: {vlookup_count}/25 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count}/25 supplier names filled")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Discount tier formulas in col H (0.15 points)
    # Initial: col H is empty. Golden: nested IF formulas based on qty.
    # Discount: 5% for 100+, 10% for 500+, 15% for 1000+
    # =========================================================================
    try:
        discount_count = 0
        for r in PO_ROWS:
            val = ws_po.cell(row=r, column=8).value  # col H = Discount Tier
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                if "IF" in val_str:
                    discount_count += 1
                else:
                    # Could be a computed value - check numeric discount
                    try:
                        qty = float(ws_po.cell(row=r, column=6).value)
                        disc_val = float(val)
                        if qty >= 1000 and abs(disc_val - 0.15) < 0.001:
                            discount_count += 1
                        elif 500 <= qty < 1000 and abs(disc_val - 0.10) < 0.001:
                            discount_count += 1
                        elif 100 <= qty < 500 and abs(disc_val - 0.05) < 0.001:
                            discount_count += 1
                        elif qty < 100 and abs(disc_val) < 0.001:
                            discount_count += 1
                    except (ValueError, TypeError):
                        pass

        ratio = discount_count / 25.0
        if ratio >= 0.9:
            print(f"PASS: Component 2 — Discount tier formulas: {discount_count}/25 (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.5:
            partial = 0.15 * ratio
            print(f"PARTIAL: Component 2 — Discount tier: {discount_count}/25 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {discount_count}/25 discount tiers filled")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Total calculation in col I (0.15 points)
    # Initial: col I is empty. Golden: =Qty*Price*(1-Discount) formulas.
    # =========================================================================
    try:
        total_count = 0
        for r in PO_ROWS:
            val = ws_po.cell(row=r, column=9).value  # col I = Total
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                # Check for formula containing multiplication pattern
                if "=" in val_str and ("*" in val_str or "F" in val_str):
                    total_count += 1
                else:
                    # Could be computed value - verify against expected
                    try:
                        qty = float(ws_po.cell(row=r, column=6).value)
                        price = float(ws_po.cell(row=r, column=7).value)
                        disc_val = ws_po.cell(row=r, column=8).value
                        if disc_val is not None:
                            try:
                                disc = float(disc_val)
                            except (ValueError, TypeError):
                                disc = 0
                        else:
                            disc = 0
                        expected_total = qty * price * (1 - disc)
                        actual = float(val)
                        if abs(actual - expected_total) < 0.1:
                            total_count += 1
                    except (ValueError, TypeError):
                        pass

        ratio = total_count / 25.0
        if ratio >= 0.9:
            print(f"PASS: Component 3 — Total formulas: {total_count}/25 (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.5:
            partial = 0.15 * ratio
            print(f"PARTIAL: Component 3 — Total formulas: {total_count}/25 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {total_count}/25 total formulas filled")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Expected delivery formulas in col J (0.15 points)
    # Initial: col J is empty. Golden: =PO_Date + terms_days formulas.
    # =========================================================================
    try:
        delivery_count = 0
        for r in PO_ROWS:
            val = ws_po.cell(row=r, column=10).value  # col J = Expected Delivery
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                # Check for formula referencing B column (PO Date) and VLOOKUP
                if "=" in val_str and ("VLOOKUP" in val_str or "B" in val_str):
                    delivery_count += 1
                else:
                    # Could be a computed date value - just check non-empty
                    import datetime
                    if isinstance(val, datetime.datetime):
                        delivery_count += 1

        ratio = delivery_count / 25.0
        if ratio >= 0.9:
            print(f"PASS: Component 4 — Expected delivery formulas: {delivery_count}/25 (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.5:
            partial = 0.15 * ratio
            print(f"PARTIAL: Component 4 — Expected delivery: {delivery_count}/25 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {delivery_count}/25 expected delivery formulas filled")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Status formulas in col L (0.15 points)
    # Initial: col L is empty. Golden: IF formulas for Delivered/Late/Pending.
    # =========================================================================
    try:
        status_count = 0
        for r in PO_ROWS:
            val = ws_po.cell(row=r, column=12).value  # col L = Status
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                if "IF" in val_str:
                    status_count += 1
                elif val_str in ['DELIVERED', 'LATE', 'PENDING']:
                    # Computed value - valid status
                    status_count += 1

        ratio = status_count / 25.0
        if ratio >= 0.9:
            print(f"PASS: Component 5 — Status formulas: {status_count}/25 (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.5:
            partial = 0.15 * ratio
            print(f"PARTIAL: Component 5 — Status formulas: {status_count}/25 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {status_count}/25 status formulas filled")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Status Board summary formulas (0.10 points)
    # Initial: Status Board counts/percentages are empty.
    # Golden: COUNTIF formulas for Delivered/Pending/Late + SUM for total.
    # =========================================================================
    try:
        sb_score = 0.0
        # Check B4 (Delivered count), B5 (Pending count), B6 (Late count), B7 (Total)
        checks = [
            (4, 2, 'COUNTIF', 'Delivered count'),
            (5, 2, 'COUNTIF', 'Pending count'),
            (6, 2, 'COUNTIF', 'Late count'),
            (7, 2, 'SUM', 'Total POs'),
        ]
        sb_pass = 0
        for row, col, expected_func, desc in checks:
            val = ws_status.cell(row=row, column=col).value
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                if expected_func.upper() in val_str:
                    sb_pass += 1
                else:
                    # Could be computed numeric value
                    try:
                        num_val = float(val)
                        if num_val >= 0:
                            sb_pass += 1
                    except (ValueError, TypeError):
                        pass

        # Also check percentage formulas in C4:C6
        pct_pass = 0
        for row in [4, 5, 6]:
            val = ws_status.cell(row=row, column=3).value
            if val is not None:
                pct_pass += 1

        # sb_pass out of 4, pct_pass out of 3
        if sb_pass >= 3 and pct_pass >= 2:
            print(f"PASS: Component 6 — Status Board formulas: counts={sb_pass}/4, pcts={pct_pass}/3 (0.10 pts)")
            total_score += 0.10
        elif sb_pass >= 2:
            partial = 0.10 * (sb_pass / 4.0)
            print(f"PARTIAL: Component 6 — Status Board: counts={sb_pass}/4, pcts={pct_pass}/3 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Status Board counts={sb_pass}/4, pcts={pct_pass}/3")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # =========================================================================
    # Component 7: Conditional formatting for Late (red) and Delivered (green) (0.10 points)
    # Initial: no conditional formatting. Golden: CF rules on L2:L26.
    # =========================================================================
    try:
        cf_rules = list(ws_po.conditional_formatting)
        late_cf_count = 0
        delivered_cf_count = 0

        for cf in cf_rules:
            cf_range = str(cf).upper()
            # Check if CF applies to column L (status column)
            if 'L' in cf_range:
                for rule in cf.rules:
                    formula_str = str(getattr(rule, 'formula', '')).upper()
                    if 'LATE' in formula_str:
                        late_cf_count += 1
                    if 'DELIVERED' in formula_str or 'DELIVER' in formula_str:
                        delivered_cf_count += 1

        if late_cf_count > 0 and delivered_cf_count > 0:
            print(f"PASS: Component 7 — Conditional formatting: Late(red)={late_cf_count}, Delivered(green)={delivered_cf_count} (0.10 pts)")
            total_score += 0.10
        elif late_cf_count > 0 or delivered_cf_count > 0:
            print(f"PARTIAL: Component 7 — CF: Late={late_cf_count}, Delivered={delivered_cf_count} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — No conditional formatting rules found for status column")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
