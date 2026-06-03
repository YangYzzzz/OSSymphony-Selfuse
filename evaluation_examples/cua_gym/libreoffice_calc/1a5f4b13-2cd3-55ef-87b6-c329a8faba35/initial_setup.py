"""
Initial Setup: Set up data validation dropdown for region filter
Task ID: calc_nrv_089
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_089'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws.cell(row=1, column=1, value="Region")
    ws.cell(row=1, column=2, value="Filter Region")

    # Populate A2:A100 with region names, many duplicates
    regions = ["North", "South", "East", "West", "Central",
               "Northeast", "Southeast", "Northwest", "Southwest", "Midwest"]

    # Create a weighted distribution so some regions appear much more often
    weights = [25, 18, 15, 12, 10, 5, 5, 4, 3, 2]  # sums to 99 rows

    region_list = []
    for region, count in zip(regions, weights):
        region_list.extend([region] * count)

    # Shuffle for realism
    random.seed(42)
    random.shuffle(region_list)

    for i, region in enumerate(region_list):
        ws.cell(row=i + 2, column=1, value=region)

    # B2 is intentionally left empty - no validation here
    # The task is for the agent to add validation

    # Set column widths for readability
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
