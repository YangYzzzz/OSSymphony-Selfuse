"""
Initial Setup: Create a Feedback sheet with reviewer data (unprotected).
Task ID: calc_ps_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Feedback'

    # --- Headers in row 1 ---
    headers = ['Reviewer', 'Score', 'Comment', 'Date']
    header_font = Font(name='Calibri', size=11, bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # --- Realistic reviewer names ---
    reviewers = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Park', 'David Rodriguez',
        'Aisha Thompson', 'James Wilson', 'Priya Patel', 'Michael Brown',
        'Lisa Nakamura', 'Robert Kim', 'Jennifer Garcia', 'Thomas Lee',
        'Maria Fernandez', 'Andrew Chang', 'Rachel Green', 'Daniel Smith',
        'Olivia Turner', 'Kevin O\'Brien', 'Natalie Cruz', 'Steven Wang',
        'Hannah Morris', 'Carlos Ruiz', 'Sophie Bennett', 'William Taylor',
        'Angela White', 'Christopher Davis', 'Megan Hill', 'Patrick Nguyen',
        'Victoria Adams', 'Brandon Clark', 'Emma Foster', 'Ryan Mitchell',
        'Laura Sanchez', 'Nathan Brooks', 'Katherine Ross', 'Jason Perez',
        'Diana Collins', 'Tyler Evans', 'Samantha Reed',
    ]

    comments = [
        'Excellent work on the quarterly analysis. Very thorough.',
        'Needs more detail in the methodology section.',
        'Good presentation skills, clear and concise delivery.',
        'The data visualizations were particularly effective.',
        'Consider adding more comparative benchmarks next time.',
        'Outstanding collaboration with the cross-functional team.',
        'Report was well-structured but could use executive summary.',
        'Great improvement from the previous review cycle.',
        'The financial projections need further validation.',
        'Solid technical foundation, recommend for promotion track.',
        'Could benefit from more stakeholder engagement.',
        'Innovative approach to solving the inventory problem.',
        'Timely delivery and high quality output overall.',
        'The risk assessment was comprehensive and well-documented.',
        'Showed strong leadership during the product launch.',
        'Needs to improve time management on deliverables.',
        'Creative problem-solving skills demonstrated throughout.',
        'The customer feedback integration was very well done.',
        'Would benefit from additional training in SQL analytics.',
        'Consistently meets deadlines with quality work.',
        'The budget analysis was accurate and actionable.',
        'Strong communication with external partners.',
        'Should focus on documenting processes more thoroughly.',
        'The competitive analysis was insightful and well-researched.',
        'Demonstrated excellent mentoring of junior team members.',
        'Report formatting and readability were top-notch.',
        'Needs to be more proactive in flagging potential issues.',
        'The testing framework proposal was highly innovative.',
        'Good attention to detail in the compliance review.',
        'The project timeline estimates were realistic and met.',
        'Should participate more actively in team discussions.',
        'The market research findings were actionable and clear.',
        'Consistently produces high-quality analytical reports.',
        'The onboarding documentation was very helpful for new hires.',
        'Could improve on cross-departmental communication.',
        'The automation initiative saved significant resources.',
        'Strong analytical skills applied to complex datasets.',
        'The presentation to the board was professional and polished.',
        'Reliable team player who consistently supports others.',
    ]

    # Dates spanning Jan 2024 to Dec 2024
    months = list(range(1, 13))

    random.seed(42)  # Reproducible data

    for i in range(39):  # rows 2 through 40
        row = i + 2
        reviewer = reviewers[i]
        score = random.randint(60, 100)
        comment = comments[i]
        month = months[i % 12]
        day = random.randint(1, 28)
        date_str = f'2024-{month:02d}-{day:02d}'

        ws.cell(row=row, column=1, value=reviewer)
        ws.cell(row=row, column=2, value=score)
        ws.cell(row=row, column=3, value=comment)
        ws.cell(row=row, column=4, value=date_str)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 14

    # Sheet is NOT protected (this is the initial state)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
