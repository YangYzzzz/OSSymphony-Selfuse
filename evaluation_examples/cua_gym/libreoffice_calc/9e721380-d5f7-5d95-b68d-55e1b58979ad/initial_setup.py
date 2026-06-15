"""
Initial Setup: Bi-weekly payroll timesheet with employee data
Task ID: calc_grs_014
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payroll"

    # --- Header area ---
    header_font = Font(name="Arial", size=14, bold=True)
    subheader_font = Font(name="Arial", size=11, bold=True)
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    white_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    white_sub_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")

    # Row 1: Company name
    ws.merge_cells("A1:U1")
    ws["A1"] = "Meridian Manufacturing Co. - Bi-Weekly Payroll Timesheet"
    ws["A1"].font = white_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2: Pay period info
    ws.merge_cells("A2:D2")
    ws["A2"] = "Pay Period: March 17, 2025 - March 30, 2025"
    ws["A2"].font = Font(name="Arial", size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 22

    ws.merge_cells("F2:I2")
    ws["F2"] = "Department: Production Floor"
    ws["F2"].font = Font(name="Arial", size=11, italic=True)

    ws.merge_cells("K2:N2")
    ws["K2"] = "Supervisor: James Rodriguez"
    ws["K2"].font = Font(name="Arial", size=11, italic=True)

    # Row 3: blank separator
    ws.row_dimensions[3].height = 8

    # --- Column headers (Row 4) ---
    col_header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers = [
        "Employee Name",   # A
        "Hourly Rate",     # B
        "Day 1\nMon",      # C
        "Day 2\nTue",      # D
        "Day 3\nWed",      # E
        "Day 4\nThu",      # F
        "Day 5\nFri",      # G
        "Day 6\nSat",      # H
        "Day 7\nSun",      # I
        "Day 8\nMon",      # J
        "Day 9\nTue",      # K
        "Day 10\nWed",     # L
        "Day 11\nThu",     # M
        "Day 12\nFri",     # N
        "Day 13\nSat",     # O
        "Day 14\nSun",     # P
        "Total Reg\nHours",   # Q
        "Total OT\nHours",    # R
        "Regular\nPay",       # S
        "Overtime\nPay",      # T
        "Gross\nPay",         # U
    ]

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[4].height = 35

    # Set column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P"]:
        ws.column_dimensions[col_letter].width = 9
    for col_letter in ["Q", "R"]:
        ws.column_dimensions[col_letter].width = 12
    for col_letter in ["S", "T", "U"]:
        ws.column_dimensions[col_letter].width = 13

    # --- Employee data (Rows 5-12) ---
    employees = [
        ("Sarah Chen",       28.50, [8.0, 8.5, 9.0, 8.0, 7.5, 4.0, 0.0, 8.0, 8.0, 10.5, 8.0, 8.0, 0.0, 0.0]),
        ("Marcus Johnson",   32.00, [8.0, 8.0, 8.0, 8.0, 8.0, 0.0, 0.0, 8.0, 9.0, 8.0, 8.0, 7.0, 0.0, 0.0]),
        ("Priya Patel",      26.75, [7.5, 8.0, 8.0, 11.0, 8.0, 6.0, 0.0, 8.0, 8.0, 8.0, 7.5, 8.0, 5.0, 0.0]),
        ("David Kowalski",   30.00, [8.0, 8.0, 9.5, 8.0, 8.0, 0.0, 0.0, 8.0, 8.0, 8.0, 12.0, 8.0, 0.0, 0.0]),
        ("Angela Torres",    27.25, [8.0, 8.0, 8.0, 8.0, 10.5, 5.0, 0.0, 8.0, 8.0, 8.0, 8.0, 8.0, 4.0, 0.0]),
        ("James Okafor",     29.00, [9.0, 8.0, 8.0, 8.0, 8.0, 0.0, 0.0, 8.0, 8.0, 11.5, 8.0, 8.0, 0.0, 0.0]),
        ("Lisa Nakamura",    31.50, [8.0, 8.0, 8.0, 8.0, 8.0, 0.0, 0.0, 8.0, 10.0, 8.0, 8.0, 8.0, 0.0, 0.0]),
        ("Robert Fitzgerald", 25.00, [8.0, 8.5, 8.0, 8.0, 8.0, 6.0, 4.0, 8.0, 8.0, 8.0, 8.0, 9.0, 5.0, 3.0]),
    ]

    light_fill_1 = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    light_fill_2 = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    name_font = Font(name="Arial", size=10, bold=True)

    for emp_idx, (name, rate, hours) in enumerate(employees):
        row = 5 + emp_idx
        row_fill = light_fill_1 if emp_idx % 2 == 0 else light_fill_2

        # Name
        cell = ws.cell(row=row, column=1, value=name)
        cell.font = name_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

        # Hourly Rate
        cell = ws.cell(row=row, column=2, value=rate)
        cell.font = data_font
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = center_align

        # Daily hours (columns C-P = 3-16)
        for day_idx, h in enumerate(hours):
            cell = ws.cell(row=row, column=3 + day_idx, value=h)
            cell.font = data_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = center_align

        # Leave columns Q-U empty (these are for formulas the agent will add)
        for col_idx in range(17, 22):
            cell = ws.cell(row=row, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border

    # --- Totals row (Row 13) - empty, to be filled by agent ---
    totals_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    totals_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")

    ws.cell(row=13, column=1, value="TOTALS").font = totals_font
    ws.cell(row=13, column=1).fill = totals_fill
    ws.cell(row=13, column=1).border = thin_border
    ws.cell(row=13, column=1).alignment = Alignment(horizontal="center", vertical="center")

    # Style rest of totals row but leave values empty
    for col_idx in range(2, 22):
        cell = ws.cell(row=13, column=col_idx)
        cell.fill = totals_fill
        cell.font = totals_font
        cell.border = thin_border
        cell.alignment = center_align

    # Freeze panes at row 4
    ws.freeze_panes = "A5"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
