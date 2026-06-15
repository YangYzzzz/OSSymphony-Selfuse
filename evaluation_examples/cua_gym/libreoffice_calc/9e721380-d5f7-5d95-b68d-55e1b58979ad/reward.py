"""
Reward Script: Bi-weekly payroll timesheet with overtime formulas, formatting, and conditional formatting
Task ID: calc_grs_014
Domain: libreoffice_calc
Scoring:
  Component 1: Regular Hours formulas in Q5:Q12 (0.20)
  Component 2: Overtime Hours formulas in R5:R12 (0.20)
  Component 3: Pay formulas in S5:U12 (Regular, OT, Gross) (0.15)
  Component 4: Totals row formulas in row 13 (0.15)
  Component 5: Number formatting (currency + hours) (0.15)
  Component 6: Conditional formatting for >10 hours (0.15)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_014'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Regular Hours formulas in Q5:Q12 (0.20 points)
    # These should be SUM(MIN(Cx,8),...) formulas computing capped-at-8 daily hours
    try:
        reg_formula_count = 0
        for row in range(5, 13):
            val = ws.cell(row=row, column=17).value  # Column Q
            if val and isinstance(val, str):
                val_norm = val.upper().replace(" ", "")
                # Check it contains MIN and 8 pattern (capping at 8 hours)
                if "MIN(" in val_norm and ",8)" in val_norm and "SUM(" in val_norm:
                    reg_formula_count += 1
        if reg_formula_count == 8:
            print(f"PASS: Component 1 — All 8 Regular Hours formulas found in Q5:Q12 (0.20 pts)")
            total_score += 0.20
        elif reg_formula_count >= 4:
            partial = 0.10
            print(f"PARTIAL: Component 1 — {reg_formula_count}/8 Regular Hours formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {reg_formula_count}/8 Regular Hours formulas found in Q5:Q12")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Overtime Hours formulas in R5:R12 (0.20 points)
    # These should be SUM(MAX(Cx-8,0),...) formulas computing hours beyond 8
    try:
        ot_formula_count = 0
        for row in range(5, 13):
            val = ws.cell(row=row, column=18).value  # Column R
            if val and isinstance(val, str):
                val_norm = val.upper().replace(" ", "")
                # Check it contains MAX and -8 pattern (overtime beyond 8)
                if "MAX(" in val_norm and "-8" in val_norm and "SUM(" in val_norm:
                    ot_formula_count += 1
        if ot_formula_count == 8:
            print(f"PASS: Component 2 — All 8 Overtime Hours formulas found in R5:R12 (0.20 pts)")
            total_score += 0.20
        elif ot_formula_count >= 4:
            partial = 0.10
            print(f"PARTIAL: Component 2 — {ot_formula_count}/8 Overtime Hours formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {ot_formula_count}/8 Overtime Hours formulas found in R5:R12")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Pay formulas in S, T, U columns for rows 5-12 (0.15 points)
    # S = Q*B (Regular Pay), T = R*B*1.5 (OT Pay), U = S+T (Gross Pay)
    try:
        pay_formula_count = 0
        for row in range(5, 13):
            # Regular Pay (column S=19): should reference Q and B
            s_val = ws.cell(row=row, column=19).value
            if s_val and isinstance(s_val, str):
                s_norm = s_val.upper().replace(" ", "")
                if f"Q{row}" in s_norm and f"B{row}" in s_norm:
                    pay_formula_count += 1

            # Overtime Pay (column T=20): should reference R, B, and 1.5
            t_val = ws.cell(row=row, column=20).value
            if t_val and isinstance(t_val, str):
                t_norm = t_val.upper().replace(" ", "")
                if f"R{row}" in t_norm and f"B{row}" in t_norm and "1.5" in t_norm:
                    pay_formula_count += 1

            # Gross Pay (column U=21): should reference S and T
            u_val = ws.cell(row=row, column=21).value
            if u_val and isinstance(u_val, str):
                u_norm = u_val.upper().replace(" ", "")
                if f"S{row}" in u_norm and f"T{row}" in u_norm:
                    pay_formula_count += 1

        # 8 employees x 3 formulas = 24 total
        if pay_formula_count >= 22:
            print(f"PASS: Component 3 — {pay_formula_count}/24 Pay formulas found in S/T/U (0.15 pts)")
            total_score += 0.15
        elif pay_formula_count >= 12:
            partial = 0.07
            print(f"PARTIAL: Component 3 — {pay_formula_count}/24 Pay formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {pay_formula_count}/24 Pay formulas found in S/T/U columns")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Totals row formulas in row 13 (0.15 points)
    # Should have SUM formulas across columns C-U in row 13
    try:
        totals_count = 0
        for col in range(3, 22):  # C=3 to U=21
            val = ws.cell(row=13, column=col).value
            if val and isinstance(val, str):
                val_norm = val.upper().replace(" ", "")
                if "SUM(" in val_norm:
                    totals_count += 1
        # Expect 19 columns (C through U) with SUM formulas
        if totals_count >= 17:
            print(f"PASS: Component 4 — {totals_count}/19 Totals row SUM formulas found (0.15 pts)")
            total_score += 0.15
        elif totals_count >= 8:
            partial = 0.07
            print(f"PARTIAL: Component 4 — {totals_count}/19 Totals row formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {totals_count}/19 Totals row formulas found in row 13")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Number formatting (0.15 points)
    # Currency format ($#,##0.00 or similar) on B,S,T,U columns
    # Hours format (0.0 or similar) on C-R columns for data rows
    try:
        currency_formatted = 0
        hours_formatted = 0

        for row in range(5, 13):
            # Check currency formatting on hourly rate (B) and pay columns (S, T, U)
            for col in [2, 19, 20, 21]:  # B, S, T, U
                fmt = ws.cell(row=row, column=col).number_format
                if fmt and ('$' in fmt or '#,##0.00' in fmt or '0.00' in fmt):
                    currency_formatted += 1

            # Check hours formatting on day columns (C-P) and totals (Q, R)
            for col in range(3, 19):  # C=3 to R=18
                fmt = ws.cell(row=row, column=col).number_format
                if fmt and fmt != 'General' and ('0.0' in fmt or '0.#' in fmt):
                    hours_formatted += 1

        # 8 employees x 4 currency cols = 32, 8 employees x 16 hour cols = 128
        currency_pct = currency_formatted / 32.0
        hours_pct = hours_formatted / 128.0

        if currency_pct >= 0.75 and hours_pct >= 0.75:
            print(f"PASS: Component 5 — Currency: {currency_formatted}/32, Hours: {hours_formatted}/128 (0.15 pts)")
            total_score += 0.15
        elif currency_pct >= 0.5 or hours_pct >= 0.5:
            partial = 0.07
            print(f"PARTIAL: Component 5 — Currency: {currency_formatted}/32, Hours: {hours_formatted}/128 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Currency: {currency_formatted}/32, Hours: {hours_formatted}/128")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting for >10 hours per day (0.15 points)
    # Should have a conditional formatting rule on the daily hours range (C5:P12)
    # that highlights cells where value > 10
    try:
        cf_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                # Check for a "greaterThan" 10 rule or formula-based >10
                is_gt10 = False
                if rule.operator == 'greaterThan' and rule.formula:
                    try:
                        threshold = float(rule.formula[0])
                        if threshold == 10:
                            is_gt10 = True
                    except (ValueError, IndexError):
                        pass

                # Also accept formula-based rules that reference >10
                if rule.type == 'expression' and rule.formula:
                    formula_str = str(rule.formula[0]) if rule.formula else ''
                    if '>10' in formula_str or '> 10' in formula_str:
                        is_gt10 = True

                if is_gt10:
                    # Check that the range covers the daily hours area (C5:P12 or similar)
                    # Accept any range that includes day columns for employee rows
                    if 'C' in cf_range and ('P' in cf_range or 'O' in cf_range or 'Q' in cf_range):
                        cf_found = True
                        break
            if cf_found:
                break

        if cf_found:
            print(f"PASS: Component 6 — Conditional formatting for >10 hours found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 6 — No conditional formatting rule for >10 hours found on daily hours range")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
