"""
Reward Script: Employee Performance Review Consolidation Sheet
Task ID: calc_gen_hr_068
Domain: libreoffice_calc
Scoring:
  - Component 1: Weighted Score formulas in Column I (0.30 pts)
  - Component 2: Rank formulas in Column J (0.20 pts)
  - Component 3: PIP Flag formulas in Column K (0.20 pts)
  - Component 4: PerfSummary sheet populated with manager summaries (0.20 pts)
  - Component 5: Conditional formatting rules present (0.10 pts)
  Total: 1.00
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_hr_068'


def normalize_formula(formula):
    """Normalize a formula string for comparison: uppercase, no spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify ReviewData sheet exists (precondition gate)
    if 'ReviewData' not in wb.sheetnames:
        print("FAIL: 'ReviewData' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws_rd = wb['ReviewData']

    # -------------------------------------------------------------------
    # Component 1: Weighted Score formula in Column I (0.30 points)
    # Formula pattern: =D*0.30+E*0.20+F*0.20+G*0.15+H*0.15
    # Must be present for all 100 data rows (I2:I101)
    # This FAILS on initial (col I is empty) and PASSES on golden
    # -------------------------------------------------------------------
    try:
        # Check header
        header_i = ws_rd.cell(row=1, column=9).value
        if header_i is None or str(header_i).strip().lower() != 'weighted score':
            print(f"FAIL: Component 1 — Column I header expected 'Weighted Score', found: {repr(header_i)}")
        else:
            # Check formulas for all 100 rows
            weighted_ok = 0
            # Expected pattern (normalized): =D*0.30+E*0.20+F*0.20+G*0.15+H*0.15
            # e.g. =D2*0.30+E2*0.20+F2*0.20+G2*0.15+H2*0.15
            formula_pattern = re.compile(
                r'^=D(\d+)\*0\.30\+E\1\*0\.20\+F\1\*0\.20\+G\1\*0\.15\+H\1\*0\.15$',
                re.IGNORECASE
            )
            for row in range(2, 102):
                cell_val = ws_rd.cell(row=row, column=9).value
                if cell_val is not None:
                    norm = normalize_formula(cell_val)
                    # Build expected pattern for this row
                    expected = f'=D{row}*0.30+E{row}*0.20+F{row}*0.20+G{row}*0.15+H{row}*0.15'.upper()
                    if norm == expected:
                        weighted_ok += 1

            if weighted_ok == 100:
                print(f"PASS: Component 1 — Weighted Score formulas present in all 100 rows I2:I101 (0.30 pts)")
                total_score += 0.30
            elif weighted_ok >= 90:
                partial = 0.20
                print(f"PARTIAL: Component 1 — Weighted Score formulas present in {weighted_ok}/100 rows, awarding {partial} pts")
                total_score += partial
            elif weighted_ok >= 50:
                partial = 0.10
                print(f"PARTIAL: Component 1 — Weighted Score formulas present in {weighted_ok}/100 rows, awarding {partial} pts")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Weighted Score formulas only in {weighted_ok}/100 rows")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Rank formula in Column J (0.20 points)
    # Formula: =RANK(I*,$I$2:$I$101)
    # Must be in all 100 data rows (J2:J101)
    # This FAILS on initial (col J is empty) and PASSES on golden
    # -------------------------------------------------------------------
    try:
        header_j = ws_rd.cell(row=1, column=10).value
        if header_j is None or str(header_j).strip().lower() != 'rank':
            print(f"FAIL: Component 2 — Column J header expected 'Rank', found: {repr(header_j)}")
        else:
            rank_ok = 0
            for row in range(2, 102):
                cell_val = ws_rd.cell(row=row, column=10).value
                if cell_val is not None:
                    norm = normalize_formula(cell_val)
                    # Expected (normalized): =RANK(I*,$I$2:$I$101)
                    expected = f'=RANK(I{row},$I$2:$I$101)'.upper()
                    if norm == expected:
                        rank_ok += 1

            if rank_ok == 100:
                print(f"PASS: Component 2 — Rank formulas present in all 100 rows J2:J101 (0.20 pts)")
                total_score += 0.20
            elif rank_ok >= 90:
                partial = 0.13
                print(f"PARTIAL: Component 2 — Rank formulas in {rank_ok}/100 rows, awarding {partial} pts")
                total_score += partial
            elif rank_ok >= 50:
                partial = 0.07
                print(f"PARTIAL: Component 2 — Rank formulas in {rank_ok}/100 rows, awarding {partial} pts")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Rank formulas only in {rank_ok}/100 rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: PIP Flag formula in Column K (0.20 points)
    # Formula: =IF(J*>=91,"PIP CANDIDATE","")
    # Must be in all 100 data rows (K2:K101)
    # This FAILS on initial (col K is empty) and PASSES on golden
    # -------------------------------------------------------------------
    try:
        header_k = ws_rd.cell(row=1, column=11).value
        if header_k is None or str(header_k).strip().lower() != 'pip flag':
            print(f"FAIL: Component 3 — Column K header expected 'PIP Flag', found: {repr(header_k)}")
        else:
            pip_ok = 0
            for row in range(2, 102):
                cell_val = ws_rd.cell(row=row, column=11).value
                if cell_val is not None:
                    norm = normalize_formula(cell_val)
                    # Expected (normalized): =IF(J*>=91,"PIP CANDIDATE","")
                    expected = f'=IF(J{row}>=91,"PIPCANDIDATE","")'.upper()
                    # Normalize the cell value too by removing spaces in strings
                    norm_stripped = norm.replace(' ', '')
                    if norm_stripped == expected:
                        pip_ok += 1

            if pip_ok == 100:
                print(f"PASS: Component 3 — PIP Flag formulas present in all 100 rows K2:K101 (0.20 pts)")
                total_score += 0.20
            elif pip_ok >= 90:
                partial = 0.13
                print(f"PARTIAL: Component 3 — PIP Flag formulas in {pip_ok}/100 rows, awarding {partial} pts")
                total_score += partial
            elif pip_ok >= 50:
                partial = 0.07
                print(f"PARTIAL: Component 3 — PIP Flag formulas in {pip_ok}/100 rows, awarding {partial} pts")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — PIP Flag formulas only in {pip_ok}/100 rows")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: PerfSummary sheet populated with manager summaries (0.20 points)
    # Headers: Manager, Avg Weighted Score, PIP Candidate Count
    # Data rows: manager names with AVERAGEIF and COUNTIFS formulas
    # This FAILS on initial (PerfSummary is empty) and PASSES on golden
    # -------------------------------------------------------------------
    try:
        if 'PerfSummary' not in wb.sheetnames:
            print("FAIL: Component 4 — 'PerfSummary' sheet not found")
        else:
            ws_ps = wb['PerfSummary']

            # Check headers
            h_a = ws_ps.cell(row=1, column=1).value
            h_b = ws_ps.cell(row=1, column=2).value
            h_c = ws_ps.cell(row=1, column=3).value

            headers_ok = (
                h_a is not None and str(h_a).strip().lower() == 'manager' and
                h_b is not None and 'avg' in str(h_b).strip().lower() and
                h_c is not None and 'pip' in str(h_c).strip().lower()
            )

            if not headers_ok:
                print(f"FAIL: Component 4 — PerfSummary headers incorrect: A1={repr(h_a)}, B1={repr(h_b)}, C1={repr(h_c)}")
            else:
                # Count data rows with manager names and formulas
                data_rows = 0
                for row in range(2, 50):
                    mgr_name = ws_ps.cell(row=row, column=1).value
                    avg_formula = ws_ps.cell(row=row, column=2).value
                    count_formula = ws_ps.cell(row=row, column=3).value
                    if mgr_name is not None:
                        # Check that avg formula contains AVERAGEIF
                        avg_ok = (avg_formula is not None and
                                  isinstance(avg_formula, str) and
                                  'AVERAGEIF' in avg_formula.upper())
                        # Check that count formula contains COUNTIFS
                        count_ok = (count_formula is not None and
                                    isinstance(count_formula, str) and
                                    'COUNTIFS' in count_formula.upper())
                        if avg_ok and count_ok:
                            data_rows += 1

                if data_rows >= 3:
                    print(f"PASS: Component 4 — PerfSummary has headers and {data_rows} manager rows with formulas (0.20 pts)")
                    total_score += 0.20
                elif data_rows >= 1:
                    partial = 0.10
                    print(f"PARTIAL: Component 4 — PerfSummary has {data_rows} complete manager rows, awarding {partial} pts")
                    total_score += partial
                else:
                    print(f"FAIL: Component 4 — PerfSummary headers OK but no complete manager data rows found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------
    # Component 5: Conditional formatting rules (0.10 points)
    # - Red fill on K2:K101 when K="PIP CANDIDATE"
    # - Gold fill on J2:J101 when J<=10
    # This FAILS on initial and PASSES on golden
    # -------------------------------------------------------------------
    try:
        cf_rules = ws_rd.conditional_formatting
        red_rule_found = False
        gold_rule_found = False

        for cf_range, rules in cf_rules._cf_rules.items():
            for rule in rules:
                if rule.type == 'expression' and rule.formula:
                    formula_str = str(rule.formula[0]).upper().replace(' ', '')
                    # Check for red fill PIP CANDIDATE rule
                    if 'PIPCANDIDATE' in formula_str or 'PIP' in formula_str:
                        # Verify it has a fill (red or any non-default)
                        if rule.dxf and rule.dxf.fill:
                            red_rule_found = True
                    # Check for gold fill on rank <=10
                    if '<=10' in formula_str:
                        if rule.dxf and rule.dxf.fill:
                            gold_rule_found = True

        if red_rule_found and gold_rule_found:
            print(f"PASS: Component 5 — Both conditional formatting rules present (red PIP + gold top-10) (0.10 pts)")
            total_score += 0.10
        elif red_rule_found or gold_rule_found:
            partial = 0.05
            which = 'red PIP rule' if red_rule_found else 'gold top-10 rule'
            print(f"PARTIAL: Component 5 — Only {which} found, awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules found on ReviewData")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
