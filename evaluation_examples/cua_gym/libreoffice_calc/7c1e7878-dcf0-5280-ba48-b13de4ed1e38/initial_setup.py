"""
Initial Setup: Employee Performance Review Consolidation Sheet
Task ID: calc_gen_hr_068
Domain: libreoffice_calc

Creates a spreadsheet with:
- ReviewData sheet: 100 employees with Emp ID, Name, Manager, and 5 competency ratings (1-5)
- PerfSummary sheet: empty (to be populated during task)

MUST NOT include: Weighted Score, Rank, PIP Flag columns or any summary data in PerfSummary
"""

import os
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_hr_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    # Realistic employee data
    first_names = [
        'Sarah', 'Marcus', 'Emily', 'James', 'Priya', 'David', 'Aisha', 'Thomas',
        'Laura', 'Kevin', 'Mei', 'Daniel', 'Fatima', 'Robert', 'Elena', 'Chris',
        'Nadia', 'Jason', 'Ingrid', 'Mohammed', 'Claire', 'Andre', 'Yuki', 'Brian',
        'Sofia', 'Ethan', 'Leila', 'Patrick', 'Amara', 'William', 'Zoe', 'Carlos',
        'Hannah', 'Tyler', 'Rania', 'Nathan', 'Vivian', 'Samuel', 'Katrina', 'Leo'
    ]
    last_names = [
        'Chen', 'Johnson', 'Williams', 'Martinez', 'Patel', 'Thompson', 'Hassan',
        'Anderson', 'Fischer', 'Park', 'Zhang', 'Brown', 'Okonkwo', 'Davis',
        'Kowalski', 'Taylor', 'Ivanova', 'Wilson', 'Lindqvist', 'Al-Rashid',
        'Dubois', 'Santos', 'Nakamura', 'Hughes', 'Reyes', 'Miller', 'Nazari',
        'O\'Brien', 'Diallo', 'Scott', 'Hoffman', 'Gomez', 'Petrov', 'Brooks',
        'Khalil', 'Carter', 'Nguyen', 'Turner', 'Johansson', 'Rivera'
    ]

    managers = [
        'Olivia Harrington', 'Derek Sutherland', 'Priscilla Watts',
        'Xavier Montenegro', 'Cecilia Park', 'Graham Foster',
        'Natasha Volkov', 'Sebastian Ruiz'
    ]

    wb = openpyxl.Workbook()

    # --- Sheet 1: ReviewData ---
    ws1 = wb.active
    ws1.title = 'ReviewData'

    # Headers
    headers = ['Emp ID', 'Name', 'Manager', 'Technical', 'Communication',
               'Leadership', 'Initiative', 'Teamwork']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Generate 100 employees
    used_names = set()
    emp_id = 1001

    for row in range(2, 102):
        # Generate unique name
        while True:
            fn = random.choice(first_names)
            ln = random.choice(last_names)
            name = f'{fn} {ln}'
            if name not in used_names:
                used_names.add(name)
                break

        manager = random.choice(managers)

        # Ratings 1-5 for 5 competencies
        technical = random.randint(1, 5)
        communication = random.randint(1, 5)
        leadership = random.randint(1, 5)
        initiative = random.randint(1, 5)
        teamwork = random.randint(1, 5)

        ws1.cell(row=row, column=1, value=f'EMP{emp_id}')
        ws1.cell(row=row, column=2, value=name)
        ws1.cell(row=row, column=3, value=manager)
        ws1.cell(row=row, column=4, value=technical)
        ws1.cell(row=row, column=5, value=communication)
        ws1.cell(row=row, column=6, value=leadership)
        ws1.cell(row=row, column=7, value=initiative)
        ws1.cell(row=row, column=8, value=teamwork)

        emp_id += 1

    # --- Sheet 2: PerfSummary (empty) ---
    ws2 = wb.create_sheet('PerfSummary')
    # Intentionally left empty per task specification

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: ReviewData (100 employees, ratings only), PerfSummary (empty)')
    print('NOTE: No Weighted Score, Rank, or PIP Flag columns in initial file')

create_initial()
