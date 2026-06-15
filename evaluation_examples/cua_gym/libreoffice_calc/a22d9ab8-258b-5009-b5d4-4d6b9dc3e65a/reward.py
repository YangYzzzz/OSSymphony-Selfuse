"""
Reward Script: Configure sheet protection to allow cell formatting but prevent data entry
Task ID: calc_gsi_090
Domain: libreoffice_calc
Scoring:
  Component 1: Sheet protection enabled on Financial Report (0.25)
  Component 2: Sheet protection enabled on Department Budget (0.25)
  Component 3: Format cells/columns/rows allowed on Financial Report (0.20)
  Component 4: Format cells/columns/rows allowed on Department Budget (0.20)
  Component 5: Cells remain locked (data entry prevented) on protected sheets (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_090'


def persist_app_state(domain: str):
    """Save any unsaved changes in LibreOffice before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Key insight: In openpyxl SheetProtection:
      - protection.sheet = True means the sheet IS protected
      - protection.formatCells = False means formatting IS ALLOWED
        (False = not restricted, True = restricted when sheet is protected)
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify expected sheets exist (precondition gate)
    expected_sheets = ['Financial Report', 'Department Budget']
    for sn in expected_sheets:
        if sn not in wb.sheetnames:
            print(f"CRITICAL: Expected sheet '{sn}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Sheet protection enabled on "Financial Report" (0.25 points)
    try:
        ws_fr = wb['Financial Report']
        if ws_fr.protection.sheet is True:
            print(f"PASS: Component 1 — Financial Report sheet protection is enabled (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Financial Report protection.sheet={ws_fr.protection.sheet}, expected True")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sheet protection enabled on "Department Budget" (0.25 points)
    try:
        ws_db = wb['Department Budget']
        if ws_db.protection.sheet is True:
            print(f"PASS: Component 2 — Department Budget sheet protection is enabled (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Department Budget protection.sheet={ws_db.protection.sheet}, expected True")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Format cells/columns/rows ALLOWED on "Financial Report" (0.20 points)
    # In openpyxl, when sheet is protected, formatCells=False means formatting IS allowed.
    try:
        ws_fr = wb['Financial Report']
        format_allowed_count = 0
        # formatCells should be False (allowed)
        if ws_fr.protection.formatCells is False:
            format_allowed_count += 1
        else:
            print(f"  INFO: Financial Report formatCells={ws_fr.protection.formatCells}, expected False (allowed)")
        # formatRows should be False (allowed)
        if ws_fr.protection.formatRows is False:
            format_allowed_count += 1
        else:
            print(f"  INFO: Financial Report formatRows={ws_fr.protection.formatRows}, expected False (allowed)")
        # formatColumns should be False (allowed)
        if ws_fr.protection.formatColumns is False:
            format_allowed_count += 1
        else:
            print(f"  INFO: Financial Report formatColumns={ws_fr.protection.formatColumns}, expected False (allowed)")

        # Only award points if sheet protection is also enabled (otherwise format permissions are meaningless)
        if ws_fr.protection.sheet is True and format_allowed_count >= 2:
            # Partial: at least formatCells must be allowed
            if ws_fr.protection.formatCells is False:
                pts = 0.20 if format_allowed_count == 3 else 0.15
                print(f"PASS: Component 3 — Financial Report formatting allowed ({format_allowed_count}/3 format permissions, {pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 3 — Financial Report formatCells is restricted (True), but task requires formatting to be allowed")
        elif ws_fr.protection.sheet is True and format_allowed_count == 1 and ws_fr.protection.formatCells is False:
            print(f"PARTIAL: Component 3 — Financial Report only formatCells allowed (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Financial Report formatting not properly configured (sheet={ws_fr.protection.sheet}, format_count={format_allowed_count})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Format cells/columns/rows ALLOWED on "Department Budget" (0.20 points)
    try:
        ws_db = wb['Department Budget']
        format_allowed_count = 0
        if ws_db.protection.formatCells is False:
            format_allowed_count += 1
        else:
            print(f"  INFO: Department Budget formatCells={ws_db.protection.formatCells}, expected False (allowed)")
        if ws_db.protection.formatRows is False:
            format_allowed_count += 1
        else:
            print(f"  INFO: Department Budget formatRows={ws_db.protection.formatRows}, expected False (allowed)")
        if ws_db.protection.formatColumns is False:
            format_allowed_count += 1
        else:
            print(f"  INFO: Department Budget formatColumns={ws_db.protection.formatColumns}, expected False (allowed)")

        if ws_db.protection.sheet is True and format_allowed_count >= 2:
            if ws_db.protection.formatCells is False:
                pts = 0.20 if format_allowed_count == 3 else 0.15
                print(f"PASS: Component 4 — Department Budget formatting allowed ({format_allowed_count}/3 format permissions, {pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 4 — Department Budget formatCells is restricted (True)")
        elif ws_db.protection.sheet is True and format_allowed_count == 1 and ws_db.protection.formatCells is False:
            print(f"PARTIAL: Component 4 — Department Budget only formatCells allowed (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Department Budget formatting not properly configured (sheet={ws_db.protection.sheet}, format_count={format_allowed_count})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Cells locked on protected sheets (data entry prevented) (0.10 points)
    # Check that cells in the data area are locked (preventing data entry when sheet is protected)
    try:
        locked_count = 0
        total_checked = 0
        for sn in expected_sheets:
            ws = wb[sn]
            if ws.protection.sheet is True:
                # Sample data cells to verify they are locked
                for row in ws.iter_rows(min_row=3, max_row=min(8, ws.max_row), min_col=1, max_col=min(6, ws.max_column)):
                    for cell in row:
                        if cell.value is not None:
                            total_checked += 1
                            if cell.protection.locked is True:
                                locked_count += 1

        if total_checked > 0:
            lock_ratio = locked_count / total_checked
            if lock_ratio >= 0.9:
                print(f"PASS: Component 5 — Data cells are locked ({locked_count}/{total_checked} = {lock_ratio:.0%}) (0.10 pts)")
                total_score += 0.10
            elif lock_ratio >= 0.5:
                print(f"PARTIAL: Component 5 — Some data cells locked ({locked_count}/{total_checked} = {lock_ratio:.0%}) (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — Most data cells are unlocked ({locked_count}/{total_checked} = {lock_ratio:.0%})")
        else:
            print(f"FAIL: Component 5 — No cells checked (sheet protection may not be enabled)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
