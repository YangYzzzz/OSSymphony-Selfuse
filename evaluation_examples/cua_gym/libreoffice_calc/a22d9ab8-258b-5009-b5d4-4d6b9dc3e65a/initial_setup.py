"""
Initial Setup: Configure sheet protection to allow cell formatting but prevent data entry
Task ID: calc_gsi_090
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Styles ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # =========================================================
    # Sheet 1: Financial Report
    # =========================================================
    ws1 = wb.active
    ws1.title = "Financial Report"

    # Title row
    ws1.merge_cells("A1:F1")
    ws1["A1"] = "FY2025 Financial Performance Report"
    ws1["A1"].font = Font(name="Calibri", size=14, bold=True, color="1F3864")
    ws1["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Headers in row 3
    fin_headers = ["Category", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Annual Total"]
    for col, h in enumerate(fin_headers, 1):
        cell = ws1.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Financial data rows
    fin_data = [
        ["Product Revenue",       1245300.00, 1389750.00, 1512400.00, 1678900.00],
        ["Service Revenue",        387200.00,  412500.00,  445800.00,  498600.00],
        ["Licensing Fees",         156000.00,  156000.00,  178500.00,  178500.00],
        ["Cost of Goods Sold",    -623400.00, -685200.00, -742100.00, -819500.00],
        ["Employee Compensation", -534000.00, -534000.00, -567000.00, -567000.00],
        ["Office & Facilities",    -89500.00,  -91200.00,  -93700.00,  -95400.00],
        ["Marketing Expenses",    -178600.00, -195400.00, -210300.00, -245700.00],
        ["R&D Investment",        -267000.00, -278500.00, -312000.00, -334800.00],
        ["Travel & Entertainment",  -34200.00,  -41800.00,  -38500.00,  -52300.00],
        ["IT Infrastructure",      -67800.00,  -72100.00,  -85400.00,  -91600.00],
        ["Legal & Compliance",     -23400.00,  -28900.00,  -31200.00,  -27500.00],
        ["Insurance",              -18700.00,  -18700.00,  -19500.00,  -19500.00],
        ["Depreciation",           -45600.00,  -45600.00,  -48200.00,  -48200.00],
    ]

    for r, row_data in enumerate(fin_data, 4):
        ws1.cell(row=r, column=1, value=row_data[0]).border = thin_border
        for c, val in enumerate(row_data[1:], 2):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
            cell.border = thin_border

    # Annual Total formulas (column F)
    for r in range(4, 4 + len(fin_data)):
        cell = ws1.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        cell.number_format = currency_fmt
        cell.border = thin_border

    # Totals row
    total_row = 4 + len(fin_data)  # row 17
    ws1.cell(row=total_row, column=1, value="Net Income").font = Font(bold=True)
    ws1.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        col_letter = openpyxl.utils.get_column_letter(c)
        cell = ws1.cell(row=total_row, column=c,
                        value=f'=SUM({col_letter}4:{col_letter}{total_row - 1})')
        cell.number_format = currency_fmt
        cell.font = Font(bold=True)
        cell.border = thin_border

    # Column widths
    ws1.column_dimensions["A"].width = 25
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col_letter].width = 16

    # Freeze header
    ws1.freeze_panes = "A4"

    # =========================================================
    # Sheet 2: Department Budget
    # =========================================================
    ws2 = wb.create_sheet("Department Budget")

    ws2.merge_cells("A1:E1")
    ws2["A1"] = "Department Budget Allocation - FY2025"
    ws2["A1"].font = Font(name="Calibri", size=13, bold=True, color="1F3864")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    dept_headers = ["Department", "Headcount", "Annual Budget", "YTD Spend", "Remaining"]
    for col, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    dept_data = [
        ["Engineering",          48, 4250000.00, 3187500.00],
        ["Product Management",   12, 1380000.00, 1035000.00],
        ["Sales",                35, 3920000.00, 2940000.00],
        ["Marketing",            22, 2150000.00, 1612500.00],
        ["Customer Support",     28, 1960000.00, 1470000.00],
        ["Human Resources",       8,  720000.00,  540000.00],
        ["Finance & Accounting", 10,  950000.00,  712500.00],
        ["Legal",                 5,  625000.00,  468750.00],
        ["Operations",           15, 1175000.00,  881250.00],
        ["Executive",             6, 1800000.00, 1350000.00],
    ]

    for r, row_data in enumerate(dept_data, 4):
        ws2.cell(row=r, column=1, value=row_data[0]).border = thin_border
        ws2.cell(row=r, column=2, value=row_data[1]).border = thin_border
        ws2.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        for c in [3, 4]:
            cell = ws2.cell(row=r, column=c, value=row_data[c - 1])
            cell.number_format = currency_fmt
            cell.border = thin_border
        # Remaining = Budget - YTD Spend
        cell = ws2.cell(row=r, column=5, value=f'=C{r}-D{r}')
        cell.number_format = currency_fmt
        cell.border = thin_border

    # Totals
    dept_total_row = 4 + len(dept_data)  # row 14
    ws2.cell(row=dept_total_row, column=1, value="Total").font = Font(bold=True)
    ws2.cell(row=dept_total_row, column=1).border = thin_border
    for c in range(2, 6):
        col_letter = openpyxl.utils.get_column_letter(c)
        cell = ws2.cell(row=dept_total_row, column=c,
                        value=f'=SUM({col_letter}4:{col_letter}{dept_total_row - 1})')
        cell.number_format = currency_fmt if c >= 3 else '0'
        cell.font = Font(bold=True)
        cell.border = thin_border

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 12
    for col_letter in ["C", "D", "E"]:
        ws2.column_dimensions[col_letter].width = 18

    ws2.freeze_panes = "A4"

    # =========================================================
    # Sheet 3: Notes
    # =========================================================
    ws3 = wb.create_sheet("Notes")
    ws3["A1"] = "Report Notes"
    ws3["A1"].font = Font(name="Calibri", size=13, bold=True)

    notes = [
        "1. All figures are in USD and rounded to the nearest cent.",
        "2. Q4 projections are based on October actuals and Nov-Dec forecasts.",
        "3. Service revenue includes maintenance contracts and consulting fees.",
        "4. R&D investment increase in H2 reflects the new product initiative.",
        "5. Marketing expenses include $120K one-time conference sponsorship in Q4.",
        "6. Headcount figures are as of December 31, 2025.",
        "7. Department budgets were approved by the board on January 15, 2025.",
        "8. YTD Spend is calculated through end of Q3 (September 30, 2025).",
    ]
    for r, note in enumerate(notes, 3):
        ws3.cell(row=r, column=1, value=note)

    ws3.column_dimensions["A"].width = 70

    # NO sheet protection applied - this is what the agent needs to configure
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
