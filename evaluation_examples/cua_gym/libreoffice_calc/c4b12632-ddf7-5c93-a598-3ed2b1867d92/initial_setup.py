"""
Initial Setup: Cross-tab revenue analysis task
Task ID: calc_gen_analysis_034
Domain: libreoffice_calc

Creates a workbook with:
- 'Transactions' sheet: 500 rows of transaction data with Trans ID, Date, Region, Category, Product, Revenue
- 'CrossTab' sheet: exists but is empty (agent must populate it)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_analysis_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Transactions ----
    ws_trans = wb.active
    ws_trans.title = 'Transactions'

    headers = ['Trans ID', 'Date', 'Region', 'Category', 'Product', 'Revenue']
    for col, h in enumerate(headers, 1):
        cell = ws_trans.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    regions = ['North', 'South', 'East', 'West']
    categories = ['Electronics', 'Apparel', 'Home', 'Food', 'Sports']

    products = {
        'Electronics': ['Laptop Pro X1', 'Wireless Headphones', 'Smart TV 55"', 'Tablet Ultra', 'Gaming Console Z'],
        'Apparel':     ['Winter Jacket', 'Running Shoes', 'Casual Dress', 'Business Suit', 'Denim Jeans'],
        'Home':        ['Coffee Maker', 'Vacuum Cleaner', 'Bed Frame Queen', 'Standing Desk', 'Air Purifier'],
        'Food':        ['Organic Granola', 'Premium Coffee Blend', 'Protein Bars (12pk)', 'Olive Oil Extra Virgin', 'Mixed Nuts 1kg'],
        'Sports':      ['Yoga Mat Pro', 'Resistance Bands Set', 'Bicycle Helmet', 'Tennis Racket', 'Swimming Goggles'],
    }

    start_date = date(2024, 1, 1)

    for i in range(500):
        row = i + 2
        trans_id = f'TXN-{10000 + i}'
        days_offset = random.randint(0, 364)
        trans_date = start_date + timedelta(days=days_offset)
        region = random.choice(regions)
        category = random.choice(categories)
        product = random.choice(products[category])
        revenue = round(random.uniform(15.50, 2850.00), 2)

        ws_trans.cell(row=row, column=1, value=trans_id)
        ws_trans.cell(row=row, column=2, value=trans_date.strftime('%Y-%m-%d'))
        ws_trans.cell(row=row, column=3, value=region)
        ws_trans.cell(row=row, column=4, value=category)
        ws_trans.cell(row=row, column=5, value=product)
        ws_trans.cell(row=row, column=6, value=revenue)

    # Set column widths for readability
    ws_trans.column_dimensions['A'].width = 14
    ws_trans.column_dimensions['B'].width = 13
    ws_trans.column_dimensions['C'].width = 10
    ws_trans.column_dimensions['D'].width = 14
    ws_trans.column_dimensions['E'].width = 28
    ws_trans.column_dimensions['F'].width = 12

    # ---- Sheet 2: CrossTab (empty — agent must populate) ----
    ws_crosstab = wb.create_sheet('CrossTab')
    # Leave completely empty — agent's task is to build the cross-tab matrix here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Transactions sheet: 500 rows with Trans ID, Date, Region, Category, Product, Revenue')
    print('  CrossTab sheet: empty (agent must build cross-tab matrix)')

create_initial()
