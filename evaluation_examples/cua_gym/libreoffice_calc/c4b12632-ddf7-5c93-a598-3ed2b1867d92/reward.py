"""
Reward Script: Build a pivot-table-style cross-tab summary from transaction dataset
Task ID: calc_gen_analysis_034
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): CrossTab headers — category names in B1:F1, region names in A2:A5
  Component 2 (0.35): SUMIFS formulas in B2:F5 — 20 cells with region+category cross-tab
  Component 3 (0.20): Totals — column totals in row 6 (B6:F6 + G6) and row totals in G2:G5
  Component 4 (0.20): Percentage formulas in B7:F10 — 20 cells as pct of grand total
  Component 5 (0.10): Conditional formatting on B2:F5 highlighting highest value cell in gold
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_analysis_034'

EXPECTED_CATEGORIES = ['Electronics', 'Apparel', 'Home', 'Food', 'Sports']
EXPECTED_REGIONS = ['North', 'South', 'East', 'West']
GOLD_COLOR = 'FFFFD700'  # gold in ARGB


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: CrossTab sheet must exist
    if 'CrossTab' not in wb.sheetnames:
        print("CRITICAL: 'CrossTab' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CrossTab']

    # -------------------------------------------------------------------------
    # Component 1: CrossTab headers — category names in B1:F1, region names in A2:A5
    # (0.15 points)
    # These FAIL on initial (CrossTab is empty) and PASS on golden.
    # -------------------------------------------------------------------------
    try:
        # Check category headers in B1:F1
        col_headers = []
        for col in range(2, 7):  # B=2 to F=6
            val = ws.cell(row=1, column=col).value
            col_headers.append(str(val).strip() if val is not None else '')

        # Check region labels in A2:A5
        row_headers = []
        for row in range(2, 6):  # rows 2-5
            val = ws.cell(row=row, column=1).value
            row_headers.append(str(val).strip() if val is not None else '')

        # Accept case-insensitive matching for headers
        col_ok = (
            len(col_headers) == 5 and
            all(col_headers[i].lower() == EXPECTED_CATEGORIES[i].lower()
                for i in range(5))
        )
        row_ok = (
            len(row_headers) == 4 and
            all(row_headers[i].lower() == EXPECTED_REGIONS[i].lower()
                for i in range(4))
        )

        if col_ok and row_ok:
            print(f"PASS: Component 1 — Headers correct: categories={col_headers}, regions={row_headers} (0.15 pts)")
            total_score += 0.15
        else:
            if not col_ok:
                print(f"FAIL: Component 1 — Category headers wrong: expected {EXPECTED_CATEGORIES}, found {col_headers}")
            if not row_ok:
                print(f"FAIL: Component 1 — Region labels wrong: expected {EXPECTED_REGIONS}, found {row_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: SUMIFS formulas in B2:F5 (4 regions x 5 categories = 20 cells)
    # (0.35 points)
    # Each cell must contain a SUMIFS formula referencing Transactions sheet.
    # These FAIL on initial (CrossTab empty) and PASS on golden.
    # -------------------------------------------------------------------------
    try:
        sumifs_count = 0
        sumifs_total = 20  # 4 rows x 5 cols

        for row in range(2, 6):  # rows 2-5
            for col in range(2, 7):  # cols B-F (2-6)
                cell_val = ws.cell(row=row, column=col).value
                if (isinstance(cell_val, str) and
                        'SUMIFS' in cell_val.upper() and
                        'Transactions' in cell_val):
                    sumifs_count += 1

        if sumifs_count == sumifs_total:
            print(f"PASS: Component 2 — All {sumifs_total} SUMIFS formulas present in B2:F5 (0.35 pts)")
            total_score += 0.35
        elif sumifs_count >= 10:
            # Partial credit: at least half the formulas present
            partial = round(0.35 * sumifs_count / sumifs_total, 4)
            print(f"PARTIAL: Component 2 — {sumifs_count}/{sumifs_total} SUMIFS formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {sumifs_count}/{sumifs_total} SUMIFS formulas found in B2:F5")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Totals — column totals in row 6 (B6:G6) and row totals in G2:G5
    # (0.20 points)
    # Must have SUM formulas in row 6 and column G for totals.
    # These FAIL on initial (CrossTab empty) and PASS on golden.
    # -------------------------------------------------------------------------
    try:
        col_total_count = 0
        row_total_count = 0

        # Check row 6 column totals (B6:F6 should be SUM(Bx:Bx), G6 grand total)
        for col in range(2, 7):  # B6:F6
            val = ws.cell(row=6, column=col).value
            if isinstance(val, str) and 'SUM' in val.upper():
                col_total_count += 1

        # Check G6 grand total
        g6 = ws.cell(row=6, column=7).value
        if isinstance(g6, str) and 'SUM' in g6.upper():
            col_total_count += 1  # count G6 as part of totals

        # Check column G row totals (G2:G5)
        for row in range(2, 6):
            val = ws.cell(row=row, column=7).value
            if isinstance(val, str) and 'SUM' in val.upper():
                row_total_count += 1

        col_totals_ok = col_total_count >= 5  # at least B6:F6 (5 cells)
        row_totals_ok = row_total_count == 4  # G2:G5 all 4 rows

        if col_totals_ok and row_totals_ok:
            print(f"PASS: Component 3 — Column totals in row 6 ({col_total_count} cells) and row totals in G2:G5 ({row_total_count} cells) (0.20 pts)")
            total_score += 0.20
        elif col_totals_ok or row_totals_ok:
            print(f"PARTIAL: Component 3 — col_totals_ok={col_totals_ok} (count={col_total_count}), row_totals_ok={row_totals_ok} (count={row_total_count}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Totals missing: col_total_count={col_total_count}, row_total_count={row_total_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Percentage formulas in B7:F10 (4 regions x 5 categories = 20 cells)
    # Each cell should reference the corresponding SUMIFS cell divided by SUM($B$2:$F$5).
    # (0.20 points)
    # These FAIL on initial (CrossTab empty) and PASS on golden.
    # -------------------------------------------------------------------------
    try:
        pct_count = 0
        pct_total = 20

        for row in range(7, 11):  # rows 7-10
            for col in range(2, 7):  # B-F (cols 2-6)
                val = ws.cell(row=row, column=col).value
                if (isinstance(val, str) and
                        'SUM($B$2:$F$5)' in val.upper().replace(' ', '') and
                        '/' in val):
                    pct_count += 1

        if pct_count == pct_total:
            print(f"PASS: Component 4 — All {pct_total} percentage formulas present in B7:F10 (0.20 pts)")
            total_score += 0.20
        elif pct_count >= 10:
            partial = round(0.20 * pct_count / pct_total, 4)
            print(f"PARTIAL: Component 4 — {pct_count}/{pct_total} percentage formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {pct_count}/{pct_total} percentage formulas in B7:F10")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting on B2:F5 highlighting highest value in gold
    # (0.10 points)
    # CF should exist on range B2:F5 with gold fill (ARGB FFFFD700).
    # These FAIL on initial (CrossTab empty) and PASS on golden.
    # -------------------------------------------------------------------------
    try:
        cf_found = False
        cf_rules = ws.conditional_formatting

        for cf in cf_rules._cf_rules:
            cf_str = str(cf)
            # Check that the CF applies to B2:F5 range
            if 'B2:F5' in cf_str:
                for rule in cf_rules._cf_rules[cf]:
                    # Check it's a formula/expression type rule with gold fill
                    if hasattr(rule, 'dxf') and rule.dxf is not None:
                        dxf = rule.dxf
                        if dxf.fill:
                            try:
                                fill_color = dxf.fill.fgColor.rgb
                                # Accept pure gold (FFD700) with or without alpha prefix
                                if 'FFD700' in fill_color.upper():
                                    cf_found = True
                                    print(f"PASS: Component 5 — CF on B2:F5 with gold fill {fill_color} found (0.10 pts)")
                                    break
                            except Exception:
                                pass
                    # Also accept if formula includes MAX (max value highlight)
                    if not cf_found and hasattr(rule, 'formula') and rule.formula:
                        formula_str = str(rule.formula)
                        if 'MAX' in formula_str.upper():
                            # Check fill even if color check failed
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                cf_found = True
                                print(f"PASS: Component 5 — CF on B2:F5 with MAX formula found (0.10 pts)")
                                break
                if cf_found:
                    break

        if cf_found:
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No gold conditional formatting found on B2:F5")
            # Debug: show all CF ranges
            for cf in cf_rules._cf_rules:
                print(f"  Found CF range: {cf}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
