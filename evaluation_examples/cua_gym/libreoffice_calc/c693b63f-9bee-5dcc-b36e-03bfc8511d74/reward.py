"""
Reward Script: Salary band classification using nested IF formulas
Task ID: calc_hr_022
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 6 cells C2:C7 contain IF formulas
  Component 2 (0.5): Formulas implement correct band logic (thresholds & band labels)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_022'

# Ground truth: salary -> expected band
EXPECTED_BANDS = {
    2: ('Alice', 45000, 'Band 1'),   # <50000
    3: ('Bob', 72000, 'Band 2'),     # 50000-79999
    4: ('Carol', 95000, 'Band 3'),   # 80000-109999
    5: ('Dan', 120000, 'Band 4'),    # >=110000
    6: ('Eve', 58000, 'Band 2'),     # 50000-79999
    7: ('Frank', 150000, 'Band 4'),  # >=110000
}


def evaluate_if_formula(formula_str, salary):
    """
    Evaluate the nested IF formula logic to determine the band.
    Parses the formula to extract thresholds and band labels,
    then applies the logic to the given salary.
    Returns the band string if the formula is correct, or None if parsing fails.
    """
    # Normalize: remove leading =, spaces
    f = formula_str.lstrip('=').replace(' ', '').upper()

    # Extract all threshold comparisons and band labels from the nested IF
    # Pattern: IF(Bx<THRESHOLD,"BAND N",...)
    # We expect 3 thresholds for 4 bands
    thresholds = []
    bands_in_formula = []

    # Find all IF(Bx<NUMBER,"BAND N" patterns
    pattern = r'IF\(B\d+<(\d+),"(BAND\s*\d+)"'
    matches = re.findall(pattern, f, re.IGNORECASE)
    for threshold, band in matches:
        thresholds.append(int(threshold))
        bands_in_formula.append(band.upper().replace(' ', ' '))

    # Find the final else band (last quoted string not part of an IF condition)
    # It's the last "BAND X" in the formula
    all_bands = re.findall(r'"(BAND\s*\d+)"', f, re.IGNORECASE)
    if all_bands:
        final_band = all_bands[-1].upper()
    else:
        return None

    if len(thresholds) < 3 or len(all_bands) < 4:
        return None

    # Apply the nested IF logic
    for i, threshold in enumerate(thresholds):
        if salary < threshold:
            return all_bands[i]
    return final_band


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Bands' not in wb.sheetnames:
        print("FAIL: Sheet 'Bands' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Bands']

    # Component 1: All 6 cells C2:C7 contain IF formulas (0.5 points)
    # Each cell with a valid IF formula earns ~0.083 points
    try:
        formula_count = 0
        formulas = {}
        for row in range(2, 8):
            cell_val = ws.cell(row=row, column=3).value
            if cell_val is not None and isinstance(cell_val, str) and '=IF(' in cell_val.upper().replace(' ', ''):
                formula_count += 1
                formulas[row] = cell_val
                print(f"PASS: C{row} contains IF formula: {cell_val}")
            else:
                print(f"FAIL: C{row} does not contain IF formula, found: {repr(cell_val)}")

        if formula_count == 6:
            print(f"PASS: Component 1 -- All 6 cells have IF formulas (0.5 pts)")
            total_score += 0.5
        elif formula_count > 0:
            partial = round(0.5 * (formula_count / 6), 2)
            print(f"PARTIAL: Component 1 -- {formula_count}/6 cells have IF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No cells contain IF formulas")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formulas produce correct band classifications (0.5 points)
    # Verify the formula logic matches the required thresholds
    try:
        correct_count = 0
        for row, (name, salary, expected_band) in EXPECTED_BANDS.items():
            if row not in formulas:
                print(f"FAIL: C{row} has no formula to evaluate")
                continue

            result = evaluate_if_formula(formulas[row], salary)
            if result is not None and result.upper().replace(' ', '') == expected_band.upper().replace(' ', ''):
                correct_count += 1
                print(f"PASS: C{row} formula correctly yields '{expected_band}' for salary {salary}")
            else:
                print(f"FAIL: C{row} formula yields '{result}' but expected '{expected_band}' for salary {salary}")

        if correct_count == 6:
            print(f"PASS: Component 2 -- All 6 formulas produce correct bands (0.5 pts)")
            total_score += 0.5
        elif correct_count > 0:
            partial = round(0.5 * (correct_count / 6), 2)
            print(f"PARTIAL: Component 2 -- {correct_count}/6 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No formulas produce correct band results")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
