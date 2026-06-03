"""
Initial Setup: Conditional formatting rules on spreadsheet - includes erroneous blue rule
Task ID: calc_tbl_027
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance"

    # Headers
    headers = ['Employee', 'Department', 'Score', 'Quarter', 'Notes']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Realistic employee data with varied scores (negative, 0-50, >50)
    data = [
        ['Sarah Chen', 'Engineering', 78, 'Q1 2025', 'Exceeded expectations'],
        ['Marcus Johnson', 'Marketing', 42, 'Q1 2025', 'Needs improvement in outreach'],
        ['Emily Rodriguez', 'Sales', -12, 'Q1 2025', 'Returned from leave, partial quarter'],
        ['David Kim', 'Engineering', 91, 'Q1 2025', 'Outstanding performance'],
        ['Jessica Taylor', 'HR', 35, 'Q1 2025', 'Satisfactory progress'],
        ['Robert Williams', 'Finance', 67, 'Q1 2025', 'Strong analytical work'],
        ['Aisha Patel', 'Sales', -5, 'Q1 2025', 'Territory reassignment impact'],
        ['Thomas Brown', 'Marketing', 55, 'Q1 2025', 'Good campaign results'],
        ['Maria Garcia', 'Engineering', 88, 'Q1 2025', 'Key contributor to release'],
        ['James Wilson', 'Finance', 23, 'Q1 2025', 'Training period adjustment'],
        ['Linda Martinez', 'HR', 0, 'Q1 2025', 'New hire onboarding focus'],
        ['Michael Lee', 'Sales', 73, 'Q1 2025', 'Consistent deal closure'],
        ['Sophie Anderson', 'Engineering', -8, 'Q1 2025', 'Project cancellation impact'],
        ['Daniel Thompson', 'Marketing', 61, 'Q1 2025', 'Creative initiatives delivered'],
        ['Rachel Moore', 'Finance', 48, 'Q1 2025', 'Approaching target'],
        ['Kevin Jackson', 'Sales', 95, 'Q1 2025', 'Top performer'],
        ['Amanda White', 'HR', 31, 'Q1 2025', 'Process improvement underway'],
        ['Christopher Davis', 'Engineering', 82, 'Q1 2025', 'Mentoring junior engineers'],
        ['Natalie Harris', 'Marketing', -3, 'Q1 2025', 'Budget overrun penalties'],
        ['Andrew Clark', 'Finance', 56, 'Q1 2025', 'Audit preparation lead'],
        ['Olivia Lewis', 'Sales', 19, 'Q1 2025', 'New territory development'],
        ['Benjamin Walker', 'Engineering', 70, 'Q1 2025', 'Feature delivery on track'],
        ['Stephanie Hall', 'HR', 44, 'Q1 2025', 'Benefits program rollout'],
        ['Patrick Allen', 'Marketing', 87, 'Q1 2025', 'Viral campaign success'],
        ['Catherine Young', 'Finance', -15, 'Q1 2025', 'Compliance issue correction'],
        ['Ryan King', 'Sales', 63, 'Q1 2025', 'Client retention improved'],
        ['Lauren Wright', 'Engineering', 50, 'Q1 2025', 'Steady contribution'],
        ['Jason Scott', 'HR', 38, 'Q1 2025', 'Recruitment pipeline building'],
        ['Michelle Green', 'Marketing', 76, 'Q1 2025', 'Brand awareness growth'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 35

    # --- Conditional Formatting Rules on C2:C30 ---
    cell_range = "C2:C30"

    # Rule 1: Red fill for negative values (< 0)
    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="lessThan", formula=["0"], fill=red_fill))

    # Rule 2: Yellow fill for values between 0 and 50 (inclusive)
    yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="between", formula=["0", "50"], fill=yellow_fill))

    # Rule 3: Green fill for values > 50
    green_fill = PatternFill(start_color="FF00B050", end_color="FF00B050", fill_type="solid")
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="greaterThan", formula=["50"], fill=green_fill))

    # Rule 4: ERRONEOUS blue rule that paints everything blue (overrides all)
    # This rule uses a broad condition that matches all numeric values
    blue_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    ws.conditional_formatting.add(cell_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["-99999"], fill=blue_fill))

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
