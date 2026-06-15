"""
Reward Script: Select sheets 'Jan', 'Feb', 'Mar' as a group, then type 'Department' in cell A1 of all three sheets.
Task ID: calc_ps_066
Domain: libreoffice_calc
Scoring:
  - Component 1: Jan!A1 contains 'Department'  (0.33 pts)
  - Component 2: Feb!A1 contains 'Department'  (0.33 pts)
  - Component 3: Mar!A1 contains 'Department'  (0.34 pts)
  Total: 1.0
"""

import os

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_066'


def persist_app_state(domain: str):
    """Best-effort save in case the file is still open in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that A1 on sheets Jan, Feb, Mar all contain 'Department',
    and that Summary!A1 was not overwritten with 'Department'.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets must exist
    required_sheets = ['Jan', 'Feb', 'Mar']
    for sn in required_sheets:
        if sn not in wb.sheetnames:
            print(f"CRITICAL: Sheet '{sn}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Jan!A1 == 'Department' (0.33 points)
    try:
        ws = wb['Jan']
        val = ws.cell(row=1, column=1).value
        if val is not None and str(val).strip() == 'Department':
            print(f"PASS: Component 1 - Jan!A1 = {val!r} (0.33 pts)")
            total_score += 0.33
        else:
            print(f"FAIL: Component 1 - Jan!A1 expected 'Department', found {val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Feb!A1 == 'Department' (0.33 points)
    try:
        ws = wb['Feb']
        val = ws.cell(row=1, column=1).value
        if val is not None and str(val).strip() == 'Department':
            print(f"PASS: Component 2 - Feb!A1 = {val!r} (0.33 pts)")
            total_score += 0.33
        else:
            print(f"FAIL: Component 2 - Feb!A1 expected 'Department', found {val!r}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Mar!A1 == 'Department' (0.34 points)
    try:
        ws = wb['Mar']
        val = ws.cell(row=1, column=1).value
        if val is not None and str(val).strip() == 'Department':
            print(f"PASS: Component 3 - Mar!A1 = {val!r} (0.34 pts)")
            total_score += 0.34
        else:
            print(f"FAIL: Component 3 - Mar!A1 expected 'Department', found {val!r}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Guard: Summary!A1 should NOT have been overwritten with 'Department'
    # This is informational only (not scored) -- the task says only Jan/Feb/Mar
    try:
        if 'Summary' in wb.sheetnames:
            ws_sum = wb['Summary']
            sum_val = ws_sum.cell(row=1, column=1).value
            if sum_val is not None and str(sum_val).strip() == 'Department':
                print(f"WARNING: Summary!A1 was overwritten with 'Department' -- task only targets Jan/Feb/Mar")
            else:
                print(f"INFO: Summary!A1 unchanged ({sum_val!r}) -- correct")
    except Exception as e:
        print(f"INFO: Could not check Summary sheet: {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
