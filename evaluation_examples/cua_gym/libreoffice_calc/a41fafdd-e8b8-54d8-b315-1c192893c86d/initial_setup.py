"""
Initial Setup: Select sheets 'Jan', 'Feb', and 'Mar' as a group, then type 'Department' in cell A1
Task ID: calc_ps_066
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Common styling
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    currency_fmt = '$#,##0.00'

    # --- Data for each month sheet ---
    # Headers start at B1 (A1 is intentionally left EMPTY for the task)
    headers = ['Employee', 'Sales', 'Expenses', 'Net']

    jan_data = [
        ['Sarah Chen', 12500, 3200, 9300],
        ['Marcus Johnson', 15800, 4100, 11700],
        ['Priya Patel', 11200, 2900, 8300],
        ['David Kim', 18400, 5200, 13200],
        ['Elena Rodriguez', 14300, 3800, 10500],
        ['James Wright', 9800, 2600, 7200],
        ['Aisha Mohammed', 16700, 4500, 12200],
        ['Robert Taylor', 13100, 3400, 9700],
        ['Lisa Nakamura', 17900, 4800, 13100],
        ['Carlos Mendez', 11600, 3100, 8500],
        ['Sophie Laurent', 14800, 3900, 10900],
        ['Wei Zhang', 19200, 5500, 13700],
    ]

    feb_data = [
        ['Sarah Chen', 13100, 3400, 9700],
        ['Marcus Johnson', 16200, 4300, 11900],
        ['Priya Patel', 10800, 2800, 8000],
        ['David Kim', 19100, 5400, 13700],
        ['Elena Rodriguez', 15000, 4000, 11000],
        ['James Wright', 10200, 2700, 7500],
        ['Aisha Mohammed', 17300, 4700, 12600],
        ['Robert Taylor', 13800, 3600, 10200],
        ['Lisa Nakamura', 18500, 5000, 13500],
        ['Carlos Mendez', 12100, 3200, 8900],
        ['Sophie Laurent', 15400, 4100, 11300],
        ['Wei Zhang', 20000, 5700, 14300],
    ]

    mar_data = [
        ['Sarah Chen', 14200, 3700, 10500],
        ['Marcus Johnson', 17000, 4500, 12500],
        ['Priya Patel', 11900, 3100, 8800],
        ['David Kim', 20300, 5800, 14500],
        ['Elena Rodriguez', 15700, 4200, 11500],
        ['James Wright', 10900, 2900, 8000],
        ['Aisha Mohammed', 18100, 4900, 13200],
        ['Robert Taylor', 14500, 3800, 10700],
        ['Lisa Nakamura', 19200, 5200, 14000],
        ['Carlos Mendez', 12800, 3400, 9400],
        ['Sophie Laurent', 16100, 4300, 11800],
        ['Wei Zhang', 21100, 6000, 15100],
    ]

    month_sheets = [
        ('Jan', jan_data),
        ('Feb', feb_data),
        ('Mar', mar_data),
    ]

    for idx, (sheet_name, data) in enumerate(month_sheets):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        # A1 is intentionally LEFT EMPTY — the task requires the agent to type 'Department' there
        # Headers in B1:E1
        for col_offset, h in enumerate(headers):
            cell = ws.cell(row=1, column=col_offset + 2, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data rows starting at row 2 (A column left empty, data in B-E)
        for r, row_data in enumerate(data, 2):
            for c, val in enumerate(row_data, 2):
                cell = ws.cell(row=r, column=c, value=val)
                if c in (3, 4, 5):  # Sales, Expenses, Net columns
                    cell.number_format = currency_fmt

        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 14
        ws.column_dimensions['D'].width = 14
        ws.column_dimensions['E'].width = 14

    # --- Summary sheet ---
    ws_summary = wb.create_sheet('Summary')
    ws_summary['A1'] = 'Quarterly Sales Summary'
    ws_summary['A1'].font = Font(name="Calibri", size=14, bold=True)

    ws_summary['A3'] = 'Month'
    ws_summary['B3'] = 'Total Sales'
    ws_summary['C3'] = 'Total Expenses'
    ws_summary['D3'] = 'Total Net'
    for col in range(1, 5):
        cell = ws_summary.cell(row=3, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    summary_data = [
        ['January', 175300, 46500, 128800],
        ['February', 181500, 48900, 132600],
        ['March', 191800, 51800, 140000],
    ]
    for r, row_data in enumerate(summary_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = currency_fmt

    ws_summary.column_dimensions['A'].width = 15
    ws_summary.column_dimensions['B'].width = 16
    ws_summary.column_dimensions['C'].width = 18
    ws_summary.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
