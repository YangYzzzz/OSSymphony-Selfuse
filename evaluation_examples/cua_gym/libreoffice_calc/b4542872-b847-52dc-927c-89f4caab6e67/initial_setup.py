"""
Initial Setup: Set zoom levels on sheets
Task ID: calc_ps_070
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Regional Sales Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers1 = ['Region', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales', 'Annual Total']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    data1 = [
        ['Northeast', 125400, 138200, 142800, 156300, 562700],
        ['Southeast', 98700, 105300, 112400, 119800, 436200],
        ['Midwest', 87300, 92100, 96500, 101200, 377100],
        ['Southwest', 76500, 81200, 85900, 90400, 334000],
        ['West Coast', 145600, 152300, 159800, 168200, 625900],
        ['Pacific Northwest', 63200, 67800, 71500, 75900, 278400],
        ['Mountain', 54800, 58200, 61700, 65300, 240000],
        ['Mid-Atlantic', 112300, 118900, 124500, 131200, 486900],
        ['New England', 71400, 76200, 80800, 85600, 314000],
        ['Great Plains', 42100, 45600, 48300, 51200, 187200],
        ['Gulf Coast', 89500, 94200, 99100, 104300, 387100],
        ['Great Lakes', 103800, 109400, 115200, 121500, 449900],
    ]

    for r, row_data in enumerate(data1, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'

    # Column widths
    ws1.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws1.column_dimensions[col_letter].width = 14

    # Zoom at default 100% (explicitly set for clarity)
    ws1.sheet_view.zoomScale = 100

    # --- Sheet2: Employee Performance ---
    ws2 = wb.create_sheet('Sheet2')

    headers2 = ['Employee', 'Department', 'Rating', 'Projects', 'Revenue Generated', 'Hire Date']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    data2 = [
        ['Sarah Chen', 'Engineering', 4.8, 12, 245000, '2022-03-15'],
        ['Marcus Johnson', 'Marketing', 4.2, 8, 189000, '2021-06-01'],
        ['Priya Patel', 'Data Science', 4.6, 10, 312000, '2023-01-10'],
        ['James Wilson', 'Sales', 4.9, 15, 478000, '2020-11-20'],
        ['Elena Rodriguez', 'Engineering', 4.5, 11, 267000, '2022-08-05'],
        ['David Kim', 'Product', 4.3, 9, 198000, '2023-04-18'],
        ['Aisha Mohammed', 'Marketing', 4.7, 13, 356000, '2021-02-14'],
        ['Robert Taylor', 'Sales', 4.1, 7, 167000, '2023-07-22'],
        ['Lisa Wang', 'Data Science', 4.4, 10, 289000, '2022-05-30'],
        ['Michael Brown', 'Engineering', 4.6, 14, 334000, '2020-09-12'],
        ['Jennifer Lee', 'Product', 4.8, 11, 298000, '2021-12-01'],
        ['Carlos Reyes', 'Sales', 4.0, 6, 145000, '2023-10-08'],
    ]

    for r, row_data in enumerate(data2, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 5:
                cell.number_format = '$#,##0'
            elif c == 3:
                cell.number_format = '0.0'

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 10
    ws2.column_dimensions['D'].width = 10
    ws2.column_dimensions['E'].width = 18
    ws2.column_dimensions['F'].width = 14

    # Zoom at default 100%
    ws2.sheet_view.zoomScale = 100

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
