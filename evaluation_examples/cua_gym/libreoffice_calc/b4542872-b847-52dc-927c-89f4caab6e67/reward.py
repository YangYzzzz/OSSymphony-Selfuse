"""
Reward Script: Set zoom levels on Sheet1 (75%) and Sheet2 (200%)
Task ID: calc_ps_070
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): Sheet1 zoom == 75%
  - Component 2 (0.5): Sheet2 zoom == 200%
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_070'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Both sheets must exist
    if 'Sheet1' not in wb.sheetnames or 'Sheet2' not in wb.sheetnames:
        print(f"FAIL: Required sheets not found. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Sheet1 zoom level == 75% (0.5 points)
    try:
        ws1 = wb['Sheet1']
        zoom1 = ws1.sheet_view.zoomScale
        if zoom1 == 75:
            print(f"PASS: Component 1 — Sheet1 zoom is 75% (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Sheet1 zoom expected 75%, found {zoom1}%")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sheet2 zoom level == 200% (0.5 points)
    try:
        ws2 = wb['Sheet2']
        zoom2 = ws2.sheet_view.zoomScale
        if zoom2 == 200:
            print(f"PASS: Component 2 — Sheet2 zoom is 200% (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — Sheet2 zoom expected 200%, found {zoom2}%")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
