"""
Reward Script: Set height of rows 11-15 to 0 to visually hide them
Task ID: calc_fmt_row_height_hide_reveal_080
Domain: libreoffice_calc
Scoring:
  Component 1: Rows 11-15 all have height == 0.0 (0.7 points)
  Component 2: Cell data in rows 11-15 is preserved unchanged (0.3 points)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_row_height_hide_reveal_080'

# Expected cell data in rows 11-15 (from task context / initial file)
EXPECTED_ROW_DATA = {
    11: {1: 'Salary Range Min'},
    12: {1: 'Salary Range Max'},
    13: {1: 'Bonus Pool'},
    14: {1: 'Equity Grants'},
    15: {1: 'Retention Bonus'},
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if 'Confidential Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Confidential Report' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Confidential Report']

    # Component 1: Rows 11-15 must all have height == 0.0 (0.7 points)
    # Task requirement: "Set the height of rows 11 through 15 to exactly 0"
    # This FAILS on initial (rows have height=15.0) and PASSES on golden (rows have height=0.0)
    try:
        rows_with_zero_height = []
        rows_with_nonzero_height = []

        for row_num in range(11, 16):
            rd = ws.row_dimensions.get(row_num)
            if rd is not None and rd.height == 0.0:
                rows_with_zero_height.append(row_num)
            else:
                actual_height = rd.height if rd is not None else 'default'
                rows_with_nonzero_height.append((row_num, actual_height))

        if len(rows_with_zero_height) == 5:
            print(f"PASS: Component 1 — All 5 rows (11-15) have height=0.0 (0.7 pts)")
            total_score += 0.7
        elif len(rows_with_zero_height) > 0:
            # Partial credit not awarded here as either all rows are set or not
            # (the task says all 5 rows must be set)
            print(f"FAIL: Component 1 — Only {len(rows_with_zero_height)}/5 rows have height=0.0")
            print(f"  Rows with height=0: {rows_with_zero_height}")
            print(f"  Rows NOT at height=0: {rows_with_nonzero_height}")
        else:
            print(f"FAIL: Component 1 — No rows (11-15) have height=0.0")
            print(f"  Rows NOT at height=0: {rows_with_nonzero_height}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cell data in rows 11-15 must be preserved (0.3 points)
    # Task requirement: "Cell values must remain unchanged (data is still there, just not visible)"
    # This checks that the data still exists despite height=0
    # This FAILS on initial because we also require rows to be zero-height (compound check)
    # We combine with the row height check: data preservation only earns points if rows are hidden
    try:
        # Check that the expected labels are still present in column A for rows 11-15
        missing_data = [
            f"Row {row_num} col {col}: expected {repr(expected_val)}, got {repr(ws.cell(row=row_num, column=col).value)}"
            for row_num, expected_cols in EXPECTED_ROW_DATA.items()
            for col, expected_val in expected_cols.items()
            if ws.cell(row=row_num, column=col).value != expected_val
        ]

        # This component is a compound check: data must be intact AND rows must be height=0
        # (We require height=0 to ensure this doesn't trivially pass on initial file)
        rows_still_zero = all(
            (ws.row_dimensions.get(r) is not None and ws.row_dimensions[r].height == 0.0)
            for r in range(11, 16)
        )

        if len(missing_data) == 0 and rows_still_zero:
            print(f"PASS: Component 2 — Cell data in rows 11-15 preserved while height=0 (0.3 pts)")
            total_score += 0.3
        elif len(missing_data) == 0 and not rows_still_zero:
            print(f"FAIL: Component 2 — Data intact but rows are not zero-height (no points)")
        else:
            print(f"FAIL: Component 2 — Cell data not preserved:")
            for msg in missing_data:
                print(f"  {msg}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
