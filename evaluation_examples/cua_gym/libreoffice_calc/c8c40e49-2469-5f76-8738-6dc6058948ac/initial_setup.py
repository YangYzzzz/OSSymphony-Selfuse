"""
Initial Setup: Set row height to 0 for rows 11-15 in a Confidential Report spreadsheet
Task ID: calc_fmt_row_height_hide_reveal_080
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_row_height_hide_reveal_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Confidential Report ---
    ws = wb.active
    ws.title = 'Confidential Report'

    # Header row styling
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column headers
    headers = ['Category', 'Q1 Value', 'Q2 Value', 'Q3 Value', 'Q4 Value', 'Annual Total', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    # Set header row height
    ws.row_dimensions[1].height = 20

    # Rows 2-10: General financial/HR data (visible and non-sensitive)
    data_rows = [
        ['Base Salaries',       1250000, 1287500, 1320000, 1345000, 5202500,  'Includes all FTE'],
        ['Benefits & Insurance',  198000,  201500,  205000,  208500,  813000,  'Health/dental/vision'],
        ['Training & Development', 42000,  45000,   48000,   51000,  186000,  'L&D programs'],
        ['Office Rent',           315000,  315000,  315000,  315000, 1260000,  'HQ + 2 satellite offices'],
        ['Software Licenses',      87500,   89200,   91000,   92800,  360500,  'SaaS subscriptions'],
        ['Travel & Expenses',      64300,   58700,   72100,   49800,  244900,  'Client-facing travel'],
        ['Marketing Budget',      175000,  195000,  165000,  220000,  755000,  'Digital + events'],
        ['IT Infrastructure',      53000,   54500,   56000,   57500,  221000,  'Servers/cloud/security'],
        ['Contractor Fees',       124000,  138000,  112000,  145000,  519000,  'Project-based contracts'],
    ]

    data_font = Font(name='Calibri', size=10)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    alt_fill = PatternFill(start_color='FFE8F0FE', end_color='FFE8F0FE', fill_type='solid')

    for r_idx, row_data in enumerate(data_rows, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = data_border
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            if c_idx in (2, 3, 4, 5, 6) and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        ws.row_dimensions[r_idx].height = 15

    # Row 10 is last data row above sensitive section — add a separator line
    ws.row_dimensions[10].height = 15

    # Rows 11-15: Sensitive compensation data (visible, default height ~12.75)
    # These are the rows that the agent must set to height=0
    sensitive_rows = [
        ['Salary Range Min',    55000,   55000,   58000,   58000,  226000, 'Band L3-L5 minimum'],
        ['Salary Range Max',   145000,  145000,  152000,  152000,  594000, 'Band L3-L5 maximum'],
        ['Bonus Pool',          87500,   92000,   96000,  105000,  380500, 'Performance-linked'],
        ['Equity Grants',       62000,   68000,   74000,   80000,  284000, 'RSU vesting schedule'],
        ['Retention Bonus',     38000,   38000,   42000,   42000,  160000, 'Key talent retention'],
    ]

    sensitive_font = Font(name='Calibri', size=10, color='9C0006')
    sensitive_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')

    for r_idx, row_data in enumerate(sensitive_rows, 11):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = sensitive_font
            cell.fill = sensitive_fill
            cell.border = data_border
            if c_idx in (2, 3, 4, 5, 6) and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        # Default height — NOT 0 (task is to set these to 0)
        ws.row_dimensions[r_idx].height = 15

    # Rows 16-20: Additional non-sensitive data
    extra_rows = [
        ['Recruitment Costs',   28500,   31200,   26800,   33400,  119900, 'Hiring & onboarding'],
        ['Payroll Processing',   8200,    8200,    8200,    8200,   32800, 'Third-party payroll'],
        ['HR System License',    4800,    4800,    4800,    4800,   19200, 'HRIS annual fee'],
        ['Compliance & Audit',  15000,   15000,   18000,   18000,   66000, 'Legal & regulatory'],
        ['Misc HR Expenses',     6700,    7200,    5900,    8100,   27900, 'Ad hoc expenses'],
    ]

    for r_idx, row_data in enumerate(extra_rows, 16):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = data_font
            cell.border = data_border
            if r_idx % 2 == 0:
                cell.fill = alt_fill
            if c_idx in (2, 3, 4, 5, 6) and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
        ws.row_dimensions[r_idx].height = 15

    # Column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Confidential Report')
    print(f'  Rows 11-15 are visible with height=15pt (task is to set these to 0)')


create_initial()
