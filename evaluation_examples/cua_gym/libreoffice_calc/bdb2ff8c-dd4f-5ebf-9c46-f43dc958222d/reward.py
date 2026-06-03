"""
Reward Script: Apply 'Accent 1' cell style to cells B2:B5
Task ID: calc_fmt_cell_style_accent_095
Domain: libreoffice_calc
Scoring:
  - Component 1: Background fill color FF4472C4 (solid) on all B2:B5 (0.5 pts)
  - Component 2: Font color white FFFFFFFF on all B2:B5 (0.3 pts)
  - Component 3: Font is bold on all B2:B5 (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_cell_style_accent_095'

# The 'Accent 1' style in LibreOffice/Excel corresponds to:
# - Background fill color: FF4472C4 (opaque blue)
# - Font color: FFFFFFFF (white)
# - Font bold: True
ACCENT1_FILL_COLOR = 'FF4472C4'
ACCENT1_FONT_COLOR = 'FFFFFFFF'

# Cells that should have the Accent 1 style applied
TARGET_CELLS = ['B2', 'B3', 'B4', 'B5']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'Executive Summary' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Executive Summary' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Executive Summary']

    # Verify cell values are unchanged (precondition gate — not scored)
    expected_values = {
        'B2': 4250000,
        'B3': 2890000,
        'B4': 1620000,
        'B5': 1100000,
    }
    values_intact = True
    for coord, expected in expected_values.items():
        actual = ws[coord].value
        if actual != expected:
            print(f"PRECONDITION FAIL: {coord} value changed — expected {expected}, found {actual}")
            values_intact = False
    if not values_intact:
        print("CRITICAL: Cell values were modified. Task requires preserving values.")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Background fill color FF4472C4 (solid) applied to all B2:B5 (0.5 points)
    # The 'Accent 1' style has a distinctive blue background (FF4472C4).
    # Initial file has no fill (patternType=None, fgColor=00000000).
    # Golden file has solid fill with fgColor=FF4472C4.
    try:
        all_fill_correct = True
        fill_details = []
        for coord in TARGET_CELLS:
            cell = ws[coord]
            try:
                fg_color = cell.fill.fgColor.rgb
                pattern = cell.fill.patternType
                is_correct = (fg_color == ACCENT1_FILL_COLOR and pattern == 'solid')
                fill_details.append(f"{coord}: fgColor={fg_color}, pattern={pattern}, correct={is_correct}")
                if not is_correct:
                    all_fill_correct = False
            except Exception as e:
                fill_details.append(f"{coord}: fill read error: {e}")
                all_fill_correct = False

        for detail in fill_details:
            print(f"  {detail}")

        if all_fill_correct:
            print(f"PASS: Component 1 — All B2:B5 have Accent 1 background fill ({ACCENT1_FILL_COLOR}, solid) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Not all B2:B5 have correct Accent 1 background fill ({ACCENT1_FILL_COLOR})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Font color is white (FFFFFFFF) on all B2:B5 (0.3 points)
    # The 'Accent 1' style sets font color to white to contrast with the blue background.
    # Initial file has no font color set (default).
    try:
        all_font_color_correct = True
        font_color_details = []
        for coord in TARGET_CELLS:
            cell = ws[coord]
            try:
                font_color = cell.font.color.rgb
                is_correct = (font_color == ACCENT1_FONT_COLOR)
                font_color_details.append(f"{coord}: font.color.rgb={font_color}, correct={is_correct}")
                if not is_correct:
                    all_font_color_correct = False
            except Exception as e:
                font_color_details.append(f"{coord}: font color read error: {e}")
                all_font_color_correct = False

        for detail in font_color_details:
            print(f"  {detail}")

        if all_font_color_correct:
            print(f"PASS: Component 2 — All B2:B5 have Accent 1 font color ({ACCENT1_FONT_COLOR}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Not all B2:B5 have correct Accent 1 font color ({ACCENT1_FONT_COLOR})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Font is bold on all B2:B5 (0.2 points)
    # The 'Accent 1' style uses bold font.
    # Initial file has bold=False for B2:B5.
    try:
        all_bold_correct = True
        bold_details = []
        for coord in TARGET_CELLS:
            cell = ws[coord]
            is_bold = cell.font.bold
            bold_details.append(f"{coord}: font.bold={is_bold}")
            if not is_bold:
                all_bold_correct = False

        for detail in bold_details:
            print(f"  {detail}")

        if all_bold_correct:
            print(f"PASS: Component 3 — All B2:B5 have bold font (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Not all B2:B5 have bold font")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
