"""
Initial Setup: Apply Accent 1 cell style to key metric values
Task ID: calc_fmt_cell_style_accent_095
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_cell_style_accent_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Executive Summary ---
    ws = wb.active
    ws.title = 'Executive Summary'

    # Header row
    ws['A1'] = 'Metric'
    ws['B1'] = 'Value'

    # Style header row
    header_font = Font(name='Calibri', size=12, bold=True)
    ws['A1'].font = header_font
    ws['B1'].font = header_font

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16

    # Data rows — metrics and values
    # NOTE: B2:B5 must NOT have Accent 1 or any special cell style
    data = [
        ('Total Revenue',   4250000),
        ('Operating Costs', 2890000),
        ('EBITDA',          1620000),
        ('Net Profit',      1100000),
    ]

    for i, (metric, value) in enumerate(data, start=2):
        ws.cell(row=i, column=1, value=metric)
        ws.cell(row=i, column=2, value=value)

    # Apply number format to B column so values look like currency
    for row in range(2, 6):
        ws.cell(row=row, column=2).number_format = '#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
