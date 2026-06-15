"""
Initial Setup: Financial model workbook for auditors (unprotected)
Task ID: calc_gsi_042
Domain: libreoffice_calc

Creates a multi-sheet financial model workbook with realistic data.
Workbook structure is NOT protected - the agent must apply protection.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    date_fmt = 'yyyy-mm-dd'
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header_row(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # --- Sheet 1: Revenue ---
    ws1 = wb.active
    ws1.title = "Revenue"

    rev_headers = ["Product Line", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Annual Total"]
    style_header_row(ws1, rev_headers)

    revenue_data = [
        ["Enterprise Software", 1245000, 1312000, 1389000, 1456000],
        ["Cloud Services", 892000, 967000, 1045000, 1123000],
        ["Consulting", 345000, 378000, 412000, 389000],
        ["Hardware Solutions", 567000, 534000, 612000, 645000],
        ["Technical Support", 234000, 248000, 256000, 271000],
        ["Training Programs", 156000, 178000, 189000, 201000],
        ["Data Analytics", 423000, 467000, 512000, 558000],
        ["Security Products", 312000, 334000, 367000, 398000],
        ["Mobile Solutions", 278000, 312000, 345000, 378000],
        ["API Services", 189000, 212000, 234000, 256000],
        ["Integration Tools", 145000, 167000, 178000, 195000],
        ["Custom Development", 523000, 489000, 534000, 567000],
    ]

    for r, row_data in enumerate(revenue_data, 2):
        ws1.cell(row=r, column=1, value=row_data[0]).border = thin_border
        for c in range(1, 5):
            cell = ws1.cell(row=r, column=c + 1, value=row_data[c])
            cell.number_format = currency_fmt
            cell.border = thin_border
        # Annual total formula
        cell = ws1.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.font = Font(bold=True)

    # Totals row
    total_row = len(revenue_data) + 2
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(c)
        cell = ws1.cell(row=total_row, column=c, value=f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
        cell.number_format = currency_fmt
        cell.font = Font(bold=True)
        cell.border = thin_border

    ws1.column_dimensions["A"].width = 22
    for col in ["B", "C", "D", "E", "F"]:
        ws1.column_dimensions[col].width = 16
    ws1.freeze_panes = "B2"

    # --- Sheet 2: Expenses ---
    ws2 = wb.create_sheet("Expenses")

    exp_headers = ["Category", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Annual Total"]
    style_header_row(ws2, exp_headers)

    expense_data = [
        ["Salaries & Wages", 1890000, 1890000, 1945000, 1945000],
        ["Employee Benefits", 472500, 472500, 486250, 486250],
        ["Office Lease", 185000, 185000, 185000, 185000],
        ["Software Licenses", 89000, 92000, 95000, 98000],
        ["Marketing & Advertising", 234000, 267000, 289000, 312000],
        ["Travel & Entertainment", 67000, 78000, 82000, 95000],
        ["Equipment & Supplies", 45000, 38000, 52000, 41000],
        ["Professional Services", 112000, 98000, 134000, 121000],
        ["Insurance", 34000, 34000, 36000, 36000],
        ["Utilities & Telecom", 28000, 29000, 30000, 31000],
        ["R&D Investment", 345000, 367000, 389000, 412000],
        ["Depreciation", 156000, 156000, 156000, 156000],
    ]

    for r, row_data in enumerate(expense_data, 2):
        ws2.cell(row=r, column=1, value=row_data[0]).border = thin_border
        for c in range(1, 5):
            cell = ws2.cell(row=r, column=c + 1, value=row_data[c])
            cell.number_format = currency_fmt
            cell.border = thin_border
        cell = ws2.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.font = Font(bold=True)

    total_row_exp = len(expense_data) + 2
    ws2.cell(row=total_row_exp, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=total_row_exp, column=1).border = thin_border
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        cell = ws2.cell(row=total_row_exp, column=c, value=f'=SUM({col_letter}2:{col_letter}{total_row_exp - 1})')
        cell.number_format = currency_fmt
        cell.font = Font(bold=True)
        cell.border = thin_border

    ws2.column_dimensions["A"].width = 24
    for col in ["B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 16
    ws2.freeze_panes = "B2"

    # --- Sheet 3: Summary ---
    ws3 = wb.create_sheet("Summary")

    summary_headers = ["Metric", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Full Year"]
    style_header_row(ws3, summary_headers)

    summary_rows = [
        "Total Revenue",
        "Total Expenses",
        "Net Income",
        "Profit Margin",
        "Revenue Growth (QoQ)",
    ]

    for r, label in enumerate(summary_rows, 2):
        ws3.cell(row=r, column=1, value=label).border = thin_border
        ws3.cell(row=r, column=1).font = Font(bold=True)

    # Revenue references
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        ws3.cell(row=2, column=c, value=f"=Revenue!{col_letter}{total_row}").number_format = currency_fmt
        ws3.cell(row=2, column=c).border = thin_border
        ws3.cell(row=3, column=c, value=f"=Expenses!{col_letter}{total_row_exp}").number_format = currency_fmt
        ws3.cell(row=3, column=c).border = thin_border
        ws3.cell(row=4, column=c, value=f"={col_letter}2-{col_letter}3").number_format = currency_fmt
        ws3.cell(row=4, column=c).border = thin_border
        ws3.cell(row=4, column=c).font = Font(bold=True)
        ws3.cell(row=5, column=c, value=f"={col_letter}4/{col_letter}2").number_format = pct_fmt
        ws3.cell(row=5, column=c).border = thin_border

    # QoQ growth (starting from Q2)
    for c in range(3, 6):
        prev_col = get_column_letter(c - 1)
        col_letter = get_column_letter(c)
        ws3.cell(row=6, column=c, value=f"=({col_letter}2-{prev_col}2)/{prev_col}2").number_format = pct_fmt
        ws3.cell(row=6, column=c).border = thin_border

    ws3.column_dimensions["A"].width = 24
    for col in ["B", "C", "D", "E", "F"]:
        ws3.column_dimensions[col].width = 16

    # --- Sheet 4: Assumptions ---
    ws4 = wb.create_sheet("Assumptions")

    assumption_headers = ["Parameter", "Value", "Notes", "Last Updated"]
    style_header_row(ws4, assumption_headers)

    assumptions = [
        ["Annual Revenue Growth Target", 0.15, "Board-approved target", "2025-01-10"],
        ["Salary Increase Rate", 0.03, "Aligned with CPI forecast", "2025-01-15"],
        ["Benefits as % of Salary", 0.25, "Industry standard", "2025-01-10"],
        ["Marketing Budget Cap", 1200000, "CEO directive", "2025-02-01"],
        ["R&D Investment Floor", 1400000, "Innovation commitment", "2025-01-20"],
        ["Office Lease Term (months)", 36, "Current contract expires Dec 2027", "2024-12-15"],
        ["Tax Rate", 0.21, "Federal corporate rate", "2025-01-05"],
        ["Discount Rate (WACC)", 0.089, "Updated by finance team", "2025-03-01"],
        ["Headcount Target EOY", 245, "HR planning", "2025-02-15"],
        ["Avg Cost per Employee", 125000, "Fully loaded", "2025-01-10"],
    ]

    for r, row_data in enumerate(assumptions, 2):
        ws4.cell(row=r, column=1, value=row_data[0]).border = thin_border
        val_cell = ws4.cell(row=r, column=2, value=row_data[1])
        val_cell.border = thin_border
        if isinstance(row_data[1], float) and row_data[1] < 1:
            val_cell.number_format = pct_fmt
        elif isinstance(row_data[1], (int, float)) and row_data[1] > 1000:
            val_cell.number_format = currency_fmt
        ws4.cell(row=r, column=3, value=row_data[2]).border = thin_border
        ws4.cell(row=r, column=4, value=row_data[3]).border = thin_border
        ws4.cell(row=r, column=4).number_format = date_fmt

    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 16
    ws4.column_dimensions["C"].width = 32
    ws4.column_dimensions["D"].width = 16

    # Workbook is NOT protected (agent must do this)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
