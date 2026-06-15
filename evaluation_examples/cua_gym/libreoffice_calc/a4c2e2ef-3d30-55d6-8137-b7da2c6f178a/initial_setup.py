"""
Initial Setup: Timesheet with broken overtime formula (treats time values as plain numbers)
Task ID: calc_tbl_077
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from datetime import time as dt_time

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Timesheet ---
    ws = wb.active
    ws.title = 'Timesheet'

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Headers
    headers = ['Employee', 'Hours Worked', 'Hourly Rate ($)', 'Overtime Pay ($)']
    col_widths = [22, 16, 18, 18]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = w

    # Employee data: name, hours_worked (as time), hourly_rate
    # Time values: 9:30 = dt_time(9,30), internally stored as fraction of day
    employees = [
        ('Sarah Chen',       dt_time(9, 30),  25.00),
        ('Marcus Johnson',   dt_time(10, 15), 28.50),
        ('Priya Patel',      dt_time(7, 45),  22.00),
        ('David Kim',        dt_time(11, 0),  32.00),
        ('Elena Rodriguez',  dt_time(8, 0),   27.50),
        ('James O\'Brien',   dt_time(9, 0),   24.00),
        ('Aisha Mohammed',   dt_time(12, 30), 35.00),
        ('Tom Nakamura',     dt_time(8, 30),  26.00),
        ('Lisa Bergstrom',   dt_time(10, 45), 30.00),
        ('Wei Zhang',        dt_time(6, 30),  21.50),
        ('Carlos Mendez',    dt_time(9, 15),  29.00),
        ('Hannah Foster',    dt_time(11, 30), 33.00),
    ]

    data_align = Alignment(horizontal='center', vertical='center')

    for r, (name, hours, rate) in enumerate(employees, 2):
        # A: Employee name
        cell_a = ws.cell(row=r, column=1, value=name)
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal='left', vertical='center')

        # B: Hours worked as time value
        cell_b = ws.cell(row=r, column=2, value=hours)
        cell_b.number_format = 'h:mm'
        cell_b.border = thin_border
        cell_b.alignment = data_align

        # C: Hourly rate
        cell_c = ws.cell(row=r, column=3, value=rate)
        cell_c.number_format = '$#,##0.00'
        cell_c.border = thin_border
        cell_c.alignment = data_align

        # D: BROKEN overtime formula
        # This formula incorrectly treats B as a plain number.
        # B2 with 9:30 is actually 0.395833... internally, not 9.5
        # So IF(B2>8,...) will always be FALSE since 0.395833 < 8
        cell_d = ws.cell(row=r, column=4, value=f'=IF(B{r}>8,(B{r}-8)*C{r}*1.5,0)')
        cell_d.number_format = '$#,##0.00'
        cell_d.border = thin_border
        cell_d.alignment = data_align

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Pay Rates ---
    ws2 = wb.create_sheet('Pay Rates')
    ws2_headers = ['Position', 'Base Rate ($)', 'Overtime Multiplier']
    for col, h in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border

    pay_data = [
        ('Junior Associate', 21.50, 1.5),
        ('Associate', 25.00, 1.5),
        ('Senior Associate', 28.50, 1.5),
        ('Lead', 32.00, 1.5),
        ('Manager', 35.00, 1.5),
    ]
    for r, (pos, base, mult) in enumerate(pay_data, 2):
        ws2.cell(row=r, column=1, value=pos).border = thin_border
        c = ws2.cell(row=r, column=2, value=base)
        c.number_format = '$#,##0.00'
        c.border = thin_border
        c = ws2.cell(row=r, column=3, value=mult)
        c.border = thin_border

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 22

    # --- Sheet 3: Notes ---
    ws3 = wb.create_sheet('Notes')
    ws3['A1'] = 'Overtime Policy'
    ws3['A1'].font = Font(size=14, bold=True)
    ws3['A3'] = 'Overtime is calculated for hours worked beyond 8 hours per day.'
    ws3['A4'] = 'Overtime rate is 1.5x the regular hourly rate.'
    ws3['A5'] = 'Time entries are recorded in HH:MM format.'
    ws3['A7'] = 'Note: The overtime formulas in the Timesheet need to be verified.'
    ws3['A7'].font = Font(italic=True, color='FF0000')
    ws3.column_dimensions['A'].width = 60

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
