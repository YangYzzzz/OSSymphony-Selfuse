"""
Reward Script: Fix overtime formula to work with time values
Task ID: calc_tbl_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4) — D2 formula uses time-extraction (HOUR/MINUTE or *24) instead of raw B2 comparison
  Component 2 (0.3) — All D-column formulas (D2:D13) use time-extraction pattern
  Component 3 (0.3) — Formulas produce correct overtime logic (>8 hrs threshold, 1.5x multiplier)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_077'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def formula_uses_time_extraction(formula):
    """
    Check if a formula correctly extracts hours from a time value.
    Valid patterns:
      - HOUR(Bn) + MINUTE(Bn)/60  (the golden approach)
      - Bn*24  (another valid approach: time fraction * 24 = decimal hours)
      - HOUR(Bn) alone (partial but acceptable if minutes handled)
    Invalid:
      - Raw Bn>8 comparison (treats time fraction as hours, always <1)
    """
    if not isinstance(formula, str):
        return False
    f_upper = formula.upper().replace(" ", "")
    # Check for HOUR() usage
    if "HOUR(" in f_upper and "MINUTE(" in f_upper:
        return True
    # Check for *24 pattern (converts time fraction to decimal hours)
    if re.search(r'B\d+\*24', f_upper):
        return True
    # Check for HOUR() even without MINUTE (partial but reasonable)
    if "HOUR(" in f_upper:
        return True
    return False


def formula_has_correct_overtime_logic(formula):
    """
    Check if the formula has the correct overtime structure:
      - Compares extracted hours > 8
      - Subtracts 8 from overtime hours
      - Multiplies by rate and 1.5
    """
    if not isinstance(formula, str):
        return False
    f_upper = formula.upper().replace(" ", "")
    # Must contain IF and threshold of 8
    if "IF(" not in f_upper:
        return False
    # Must reference the 1.5 multiplier
    if "1.5" not in f_upper:
        return False
    # Must subtract 8 for overtime hours
    if "-8)" not in f_upper and "-8]*" not in f_upper and "-8)*" not in f_upper:
        return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Timesheet sheet must exist
    if 'Timesheet' not in wb.sheetnames:
        print("CRITICAL: 'Timesheet' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timesheet']

    # =========================================================================
    # Component 1: D2 formula uses time-extraction (0.4 points)
    # The task specifically mentions D2 as the formula to fix.
    # Initial: =IF(B2>8,(B2-8)*C2*1.5,0) -- treats time as raw number
    # Expected: uses HOUR()/MINUTE() or *24 to extract decimal hours
    # =========================================================================
    try:
        d2_formula = ws['D2'].value
        if d2_formula and isinstance(d2_formula, str) and d2_formula.startswith('='):
            if formula_uses_time_extraction(d2_formula):
                print(f"PASS: Component 1 — D2 formula uses time extraction: {d2_formula} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — D2 formula does NOT extract time properly: {d2_formula}")
        else:
            print(f"FAIL: Component 1 — D2 does not contain a formula: {d2_formula!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: All D-column formulas (D2:D13) use time extraction (0.3 points)
    # The fix should be applied consistently to all rows, not just D2.
    # =========================================================================
    try:
        fixed_count = 0
        total_formulas = 0
        for row in range(2, 14):  # D2 through D13
            cell_val = ws.cell(row=row, column=4).value
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                total_formulas += 1
                if formula_uses_time_extraction(cell_val):
                    fixed_count += 1
                else:
                    print(f"  D{row} still uses raw comparison: {cell_val}")

        if total_formulas > 0 and fixed_count == total_formulas:
            print(f"PASS: Component 2 — All {total_formulas} D-column formulas use time extraction (0.3 pts)")
            total_score += 0.3
        elif total_formulas > 0:
            print(f"FAIL: Component 2 — Only {fixed_count}/{total_formulas} formulas use time extraction")
        else:
            print(f"FAIL: Component 2 — No formulas found in D2:D13")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Formulas are BOTH time-aware AND have correct overtime logic (0.3 points)
    # Must: use time extraction AND compare hours > 8, subtract 8, multiply by rate * 1.5
    # This is a compound check: the overtime logic is only correct when combined with
    # proper time extraction. The initial file has the same arithmetic structure but
    # applied to raw time fractions, so this must require BOTH conditions.
    # =========================================================================
    try:
        fully_correct_count = 0
        total_checked = 0
        for row in range(2, 14):
            cell_val = ws.cell(row=row, column=4).value
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                total_checked += 1
                if formula_uses_time_extraction(cell_val) and formula_has_correct_overtime_logic(cell_val):
                    fully_correct_count += 1
                else:
                    print(f"  D{row} missing time extraction or overtime logic: {cell_val}")

        if total_checked > 0 and fully_correct_count == total_checked:
            print(f"PASS: Component 3 — All {total_checked} formulas are fully correct (time-aware + overtime logic) (0.3 pts)")
            total_score += 0.3
        elif total_checked > 0:
            print(f"FAIL: Component 3 — Only {fully_correct_count}/{total_checked} formulas fully correct")
        else:
            print(f"FAIL: Component 3 — No formulas found in D2:D13")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
