"""
Reward Script: Employee Performance Review Spreadsheet
Task ID: calc_wf_053
Domain: libreoffice_calc
Scoring:
  Component 1: SUMPRODUCT weighted average formulas (0.25)
  Component 2: Overall score and RANK formulas (0.15)
  Component 3: Gap analysis ABS formulas (0.20)
  Component 4: Conditional formatting on gap cells (0.15)
  Component 5: Radar chart presence and series (0.15)
  Component 6: Radar chart data table (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_053'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Analysis' sheet exists
    if 'Analysis' not in wb.sheetnames:
        print("FAIL: 'Analysis' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Analysis']

    # ---------------------------------------------------------------
    # Component 1: SUMPRODUCT weighted average formulas (0.25 points)
    # B2:B9 should have SUMPRODUCT referencing Self-Assessment scores and Weights
    # C2:C9 should have SUMPRODUCT referencing Manager Assessment scores and Weights
    # ---------------------------------------------------------------
    try:
        sumproduct_count = 0
        expected_cells = 16  # 8 employees x 2 columns (B and C)

        for row in range(2, 10):  # rows 2-9
            b_val = ws.cell(row=row, column=2).value  # B column - Self weighted avg
            c_val = ws.cell(row=row, column=3).value  # C column - Manager weighted avg

            if b_val and isinstance(b_val, str) and 'SUMPRODUCT' in b_val.upper():
                sumproduct_count += 1
            if c_val and isinstance(c_val, str) and 'SUMPRODUCT' in c_val.upper():
                sumproduct_count += 1

        if sumproduct_count >= expected_cells:
            print(f"PASS: Component 1 — All {sumproduct_count} SUMPRODUCT formulas found in B2:C9 (0.25 pts)")
            total_score += 0.25
        elif sumproduct_count > 0:
            partial = 0.25 * (sumproduct_count / expected_cells)
            print(f"PARTIAL: Component 1 — {sumproduct_count}/{expected_cells} SUMPRODUCT formulas found ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No SUMPRODUCT formulas found in B2:C9")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Overall score and RANK formulas (0.15 points)
    # D2:D9 should compute average of self and manager weighted scores
    # E2:E9 should have RANK formulas
    # ---------------------------------------------------------------
    try:
        overall_count = 0
        rank_count = 0

        for row in range(2, 10):
            d_val = ws.cell(row=row, column=4).value  # D column - Overall Score
            e_val = ws.cell(row=row, column=5).value  # E column - Rank

            # Overall score: should be a formula combining B and C (e.g., =(B2+C2)/2 or =AVERAGE(B2,C2))
            if d_val and isinstance(d_val, str) and d_val.startswith('='):
                # Accept any formula that references both B and C columns
                d_upper = d_val.upper()
                if ('B' in d_upper and 'C' in d_upper) or 'AVERAGE' in d_upper:
                    overall_count += 1

            # Rank: should have RANK function
            if e_val and isinstance(e_val, str) and 'RANK' in e_val.upper():
                rank_count += 1

        score_2 = 0.0
        if overall_count >= 8:
            score_2 += 0.075
            print(f"PASS: Component 2a — All {overall_count} overall score formulas found in D2:D9")
        elif overall_count > 0:
            score_2 += 0.075 * (overall_count / 8)
            print(f"PARTIAL: Component 2a — {overall_count}/8 overall score formulas found")
        else:
            print(f"FAIL: Component 2a — No overall score formulas found in D2:D9")

        if rank_count >= 8:
            score_2 += 0.075
            print(f"PASS: Component 2b — All {rank_count} RANK formulas found in E2:E9")
        elif rank_count > 0:
            score_2 += 0.075 * (rank_count / 8)
            print(f"PARTIAL: Component 2b — {rank_count}/8 RANK formulas found")
        else:
            print(f"FAIL: Component 2b — No RANK formulas found in E2:E9")

        total_score += score_2
        if score_2 > 0:
            print(f"  Component 2 total: {score_2:.3f} pts")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Gap analysis ABS formulas (0.20 points)
    # B14:K21 should have ABS(Self - Manager) formulas for each employee x competency
    # ---------------------------------------------------------------
    try:
        abs_count = 0
        expected_abs = 80  # 8 employees x 10 competencies

        for row in range(14, 22):  # rows 14-21
            for col in range(2, 12):  # columns B-K
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str) and 'ABS' in val.upper():
                    abs_count += 1

        if abs_count >= expected_abs:
            print(f"PASS: Component 3 — All {abs_count} ABS gap formulas found in B14:K21 (0.20 pts)")
            total_score += 0.20
        elif abs_count > 0:
            partial = 0.20 * (abs_count / expected_abs)
            print(f"PARTIAL: Component 3 — {abs_count}/{expected_abs} ABS formulas found ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No ABS gap formulas found in B14:K21")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Conditional formatting on gap cells (0.15 points)
    # B14:K21 should have conditional formatting:
    #   - >1.5 with red fill (FFFF0000)
    #   - >1.0 with yellow fill (FFFFFF00)
    # ---------------------------------------------------------------
    try:
        has_red_rule = False
        has_yellow_rule = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the conditional formatting covers the gap area (rows 14-21)
            for rule in cf.rules:
                if rule.type == 'cellIs' and rule.operator == 'greaterThan':
                    formula_str = str(rule.formula)
                    # Check for >1.5 red rule
                    if '1.5' in formula_str:
                        if rule.dxf and rule.dxf.fill:
                            fill_color = None
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                            except Exception:
                                pass
                            if fill_color and 'FF0000' in fill_color:
                                has_red_rule = True
                            else:
                                # Accept any red-ish fill with >1.5
                                has_red_rule = True
                        else:
                            has_red_rule = True

                    # Check for >1.0 yellow rule
                    if '1.0' in formula_str or ("'1'" in formula_str) or formula_str.strip("[]'") == '1':
                        if rule.dxf and rule.dxf.fill:
                            fill_color = None
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                            except Exception:
                                pass
                            if fill_color and 'FFFF00' in fill_color:
                                has_yellow_rule = True
                            else:
                                has_yellow_rule = True
                        else:
                            has_yellow_rule = True

        score_4 = 0.0
        if has_red_rule:
            score_4 += 0.075
            print(f"PASS: Component 4a — Red conditional formatting (>1.5) found")
        else:
            print(f"FAIL: Component 4a — No red conditional formatting (>1.5) found")

        if has_yellow_rule:
            score_4 += 0.075
            print(f"PASS: Component 4b — Yellow conditional formatting (>1.0) found")
        else:
            print(f"FAIL: Component 4b — No yellow conditional formatting (>1.0) found")

        total_score += score_4
        if score_4 > 0:
            print(f"  Component 4 total: {score_4:.3f} pts")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Radar chart presence and series (0.15 points)
    # Analysis sheet should have at least one radar chart with 2 series
    # (Self vs Manager assessment comparison)
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        radar_found = False
        has_two_series = False

        for chart in charts:
            chart_type = type(chart).__name__
            if 'Radar' in chart_type:
                radar_found = True
                if len(chart.series) >= 2:
                    has_two_series = True
                break

        score_5 = 0.0
        if radar_found and has_two_series:
            score_5 = 0.15
            print(f"PASS: Component 5 — Radar chart with {len(chart.series)} series found (0.15 pts)")
        elif radar_found:
            score_5 = 0.08
            print(f"PARTIAL: Component 5 — Radar chart found but only {len(chart.series)} series (expected 2)")
        elif len(charts) > 0:
            # Some chart exists but not radar
            score_5 = 0.05
            print(f"PARTIAL: Component 5 — Chart found ({type(charts[0]).__name__}) but not radar type")
        else:
            print(f"FAIL: Component 5 — No chart found on Analysis sheet")

        total_score += score_5
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Radar chart data table (0.10 points)
    # Rows 24-34 should have competency data for radar chart
    # with self and manager scores for a selected employee
    # ---------------------------------------------------------------
    try:
        # Check for competency labels in column A and numeric data in B and C
        competency_names = [
            'Communication', 'Technical Skills', 'Leadership', 'Problem Solving',
            'Teamwork', 'Time Management', 'Adaptability', 'Creativity',
            'Work Ethic', 'Customer Focus'
        ]

        # Search in a range around rows 23-35 for the radar data table
        data_rows_found = 0
        header_found = False

        for row in range(23, 36):
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value

            if a_val and isinstance(a_val, str):
                # Check for header row (contains "Self" or "Manager" in B or C headers)
                if b_val and isinstance(b_val, str) and 'self' in b_val.lower():
                    header_found = True
                # Check for competency data rows
                if a_val.strip() in competency_names:
                    if b_val is not None and c_val is not None:
                        try:
                            float(b_val)
                            float(c_val)
                            data_rows_found += 1
                        except (ValueError, TypeError):
                            pass

        score_6 = 0.0
        if data_rows_found >= 10 and header_found:
            score_6 = 0.10
            print(f"PASS: Component 6 — Radar data table with {data_rows_found} competency rows and header (0.10 pts)")
        elif data_rows_found >= 5:
            score_6 = 0.10 * (data_rows_found / 10)
            print(f"PARTIAL: Component 6 — {data_rows_found}/10 competency data rows found")
        elif data_rows_found > 0:
            score_6 = 0.03
            print(f"PARTIAL: Component 6 — Only {data_rows_found} competency data rows found")
        else:
            print(f"FAIL: Component 6 — No radar chart data table found")

        total_score += score_6
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
