"""
Initial Setup: Employee Performance Review Spreadsheet
Task ID: calc_wf_053
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

EMPLOYEES = [
    'Sarah Chen',
    'Marcus Johnson',
    'Priya Patel',
    'James Wilson',
    'Elena Rodriguez',
    'David Kim',
    'Rachel Thompson',
    'Ahmed Hassan',
]

COMPETENCIES = [
    'Communication',
    'Technical Skills',
    'Leadership',
    'Problem Solving',
    'Teamwork',
    'Time Management',
    'Adaptability',
    'Creativity',
    'Work Ethic',
    'Customer Focus',
]

WEIGHTS = [0.15, 0.15, 0.10, 0.12, 0.10, 0.08, 0.08, 0.07, 0.08, 0.07]

# Self-assessment scores (1-5) for each employee x competency
SELF_SCORES = [
    [4, 5, 3, 4, 5, 4, 3, 4, 5, 4],  # Sarah Chen
    [3, 4, 4, 3, 4, 3, 4, 3, 4, 3],  # Marcus Johnson
    [5, 4, 3, 5, 4, 5, 4, 5, 4, 5],  # Priya Patel
    [3, 3, 5, 4, 3, 4, 3, 3, 4, 3],  # James Wilson
    [4, 5, 4, 4, 5, 3, 5, 4, 3, 4],  # Elena Rodriguez
    [4, 4, 3, 5, 4, 4, 3, 4, 5, 4],  # David Kim
    [3, 3, 4, 3, 4, 5, 4, 3, 4, 5],  # Rachel Thompson
    [5, 4, 4, 4, 3, 3, 5, 5, 4, 3],  # Ahmed Hassan
]

# Manager assessment scores (1-5)
MANAGER_SCORES = [
    [3, 5, 3, 4, 4, 3, 3, 3, 5, 4],  # Sarah Chen
    [4, 3, 3, 4, 4, 4, 3, 3, 3, 4],  # Marcus Johnson
    [4, 5, 3, 4, 3, 4, 4, 4, 4, 4],  # Priya Patel
    [3, 4, 4, 3, 3, 3, 4, 2, 4, 3],  # James Wilson
    [3, 4, 4, 5, 4, 3, 4, 3, 4, 3],  # Elena Rodriguez
    [5, 3, 4, 4, 4, 3, 3, 4, 4, 5],  # David Kim
    [4, 4, 3, 3, 3, 4, 3, 4, 3, 4],  # Rachel Thompson
    [4, 4, 5, 3, 4, 4, 4, 4, 3, 4],  # Ahmed Hassan
]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # --- Sheet 1: Self-Assessment ---
    ws_self = wb.active
    ws_self.title = 'Self-Assessment'

    headers_self = ['Employee'] + COMPETENCIES
    for col, h in enumerate(headers_self, 1):
        cell = ws_self.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r, emp in enumerate(EMPLOYEES, 2):
        ws_self.cell(row=r, column=1, value=emp).border = thin_border
        for c, score in enumerate(SELF_SCORES[r - 2], 2):
            cell = ws_self.cell(row=r, column=c, value=score)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    ws_self.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws_self.column_dimensions[col_letter].width = 16

    # --- Sheet 2: Manager Assessment ---
    ws_mgr = wb.create_sheet('Manager Assessment')

    headers_mgr = ['Employee'] + COMPETENCIES
    for col, h in enumerate(headers_mgr, 1):
        cell = ws_mgr.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r, emp in enumerate(EMPLOYEES, 2):
        ws_mgr.cell(row=r, column=1, value=emp).border = thin_border
        for c, score in enumerate(MANAGER_SCORES[r - 2], 2):
            cell = ws_mgr.cell(row=r, column=c, value=score)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    ws_mgr.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws_mgr.column_dimensions[col_letter].width = 16

    # --- Sheet 3: Weights ---
    ws_weights = wb.create_sheet('Weights')

    weight_headers = ['Competency', 'Weight']
    for col, h in enumerate(weight_headers, 1):
        cell = ws_weights.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r, (comp, wt) in enumerate(zip(COMPETENCIES, WEIGHTS), 2):
        ws_weights.cell(row=r, column=1, value=comp).border = thin_border
        cell = ws_weights.cell(row=r, column=2, value=wt)
        cell.number_format = '0%'
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Total row
    total_cell = ws_weights.cell(row=12, column=1, value='Total')
    total_cell.font = Font(bold=True)
    total_cell.border = thin_border
    total_val = ws_weights.cell(row=12, column=2, value=1.0)
    total_val.number_format = '0%'
    total_val.font = Font(bold=True)
    total_val.alignment = Alignment(horizontal='center')
    total_val.border = thin_border

    ws_weights.column_dimensions['A'].width = 22
    ws_weights.column_dimensions['B'].width = 12

    # --- Sheet 4: Analysis (empty template - no formulas, no data, no chart) ---
    ws_analysis = wb.create_sheet('Analysis')

    analysis_headers = [
        'Employee',
        'Self Weighted Avg',
        'Manager Weighted Avg',
        'Overall Score',
        'Rank',
    ]
    for col, h in enumerate(analysis_headers, 1):
        cell = ws_analysis.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Employee names only - no formulas
    for r, emp in enumerate(EMPLOYEES, 2):
        ws_analysis.cell(row=r, column=1, value=emp).border = thin_border

    # Gap sub-header section (row 12+)
    gap_start_row = 12
    ws_analysis.cell(row=gap_start_row, column=1, value='Gap Analysis (|Self - Manager| per Competency)').font = Font(bold=True, size=12)
    ws_analysis.merge_cells(start_row=gap_start_row, start_column=1, end_row=gap_start_row, end_column=11)

    gap_headers = ['Employee'] + COMPETENCIES
    for col, h in enumerate(gap_headers, 1):
        cell = ws_analysis.cell(row=gap_start_row + 1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    for r, emp in enumerate(EMPLOYEES, gap_start_row + 2):
        ws_analysis.cell(row=r, column=1, value=emp).border = thin_border

    ws_analysis.column_dimensions['A'].width = 22
    ws_analysis.column_dimensions['B'].width = 18
    ws_analysis.column_dimensions['C'].width = 20
    ws_analysis.column_dimensions['D'].width = 16
    ws_analysis.column_dimensions['E'].width = 10
    for col_letter in ['F', 'G', 'H', 'I', 'J', 'K']:
        ws_analysis.column_dimensions[col_letter].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
