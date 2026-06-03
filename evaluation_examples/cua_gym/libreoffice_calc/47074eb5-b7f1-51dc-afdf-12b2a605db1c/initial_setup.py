"""
Initial Setup: Budget Template spreadsheet with one sheet
Task ID: calc_sht_copy_005
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_copy_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Budget Template ---
    ws = wb.active
    ws.title = 'Budget Template'

    # Row 1: Title - DEPARTMENT BUDGET 2025 (merged A1:G1)
    ws.merge_cells('A1:G1')
    ws['A1'] = 'DEPARTMENT BUDGET 2025'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFFFF')
    ws.row_dimensions[1].height = 30

    # Row 2: Department label
    ws['A2'] = 'Department:'
    ws['A2'].font = Font(name='Calibri', size=11, bold=True)
    ws['B2'] = ''  # blank for input

    # Row 3: blank spacer
    ws.row_dimensions[3].height = 10

    # Row 4: Column headers
    headers = ['Line Item', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total', 'Prior Year']
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    ws.row_dimensions[4].height = 20

    # Column widths
    ws.column_dimensions['A'].width = 30
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 14

    # Budget line items - Rows 5 to 30 (26 line items)
    line_items = [
        ('Salaries & Wages',          45000, 45000, 46000, 46000, 182000),
        ('Benefits & Insurance',       9500,  9500,  9700,  9700,  38400),
        ('Contractor / Consulting',    5000,  7500,  5000,  7500,  25000),
        ('Office Supplies',             800,   750,   800,   800,   3150),
        ('Software Licenses',          3200,  3200,  3200,  3200,  12800),
        ('Hardware & Equipment',       6000,   500,   500,   500,   7500),
        ('Travel & Expenses',          2000,  3000,  2000,  3000,  10000),
        ('Training & Development',     1500,  1500,  1500,  1500,   6000),
        ('Marketing & Advertising',    4000,  5000,  4000,  6000,  19000),
        ('Rent & Facilities',          8500,  8500,  8500,  8500,  34000),
        ('Utilities',                  1200,  1200,  1200,  1200,   4800),
        ('Telecommunications',          900,   900,   900,   900,   3600),
        ('Postage & Shipping',          300,   300,   300,   300,   1200),
        ('Printing & Reproduction',     400,   400,   400,   400,   1600),
        ('Legal & Professional Fees',  2500,  2500,  2500,  2500,  10000),
        ('Accounting & Audit',         1800,     0,  1800,     0,   3600),
        ('Insurance - General',        3000,  3000,  3000,  3000,  12000),
        ('Depreciation',               1500,  1500,  1500,  1500,   6000),
        ('Research & Development',     4500,  4500,  4500,  4500,  18000),
        ('Miscellaneous',               500,   500,   500,   500,   2000),
        ('Recruiting & Hiring',        2000,  2000,  1000,  1000,   6000),
        ('Employee Recognition',        500,   500,   500,   500,   2000),
        ('Conference & Events',        1000,  2000,  1000,  2000,   6000),
        ('Subscriptions & Memberships', 600,   600,   600,   600,   2400),
        ('IT Support & Maintenance',   2200,  2200,  2200,  2200,   8800),
        ('Contingency Reserve',        2000,  2000,  2000,  2000,   8000),
    ]

    # Prior year data (slightly different from current year totals)
    prior_year_data = [
        175000, 37000, 22000, 3000, 12000, 5000, 9500, 5500,
        17000, 32000, 4600, 3400, 1100, 1500, 9500, 3500,
        11500, 5800, 16000, 1800, 5500, 1800, 5500, 2200,
        8200, 7500
    ]

    data_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )

    for row_idx, (item, q1, q2, q3, q4, annual, *_) in enumerate(line_items, 5):
        prior = prior_year_data[row_idx - 5]
        ws.cell(row=row_idx, column=1, value=item).border = data_border
        ws.cell(row=row_idx, column=2, value=q1).border = data_border
        ws.cell(row=row_idx, column=3, value=q2).border = data_border
        ws.cell(row=row_idx, column=4, value=q3).border = data_border
        ws.cell(row=row_idx, column=5, value=q4).border = data_border
        ws.cell(row=row_idx, column=6, value=annual).border = data_border
        ws.cell(row=row_idx, column=7, value=prior).border = data_border

        # Number format for numeric columns
        for col in range(2, 8):
            ws.cell(row=row_idx, column=col).number_format = '#,##0'

        # Alternating row color
        if (row_idx - 5) % 2 == 1:
            light_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
            for col in range(1, 8):
                ws.cell(row=row_idx, column=col).fill = light_fill

    # Row 31: Totals row with SUM formulas
    total_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    total_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    total_border = Border(
        left=thin, right=thin,
        top=Side(style='medium', color='000000'),
        bottom=Side(style='medium', color='000000')
    )

    ws.cell(row=31, column=1, value='TOTAL').font = total_font
    ws.cell(row=31, column=1).fill = total_fill
    ws.cell(row=31, column=1).border = total_border

    sum_cols = {
        2: 'B', 3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G'
    }
    for col_num, col_letter in sum_cols.items():
        cell = ws.cell(row=31, column=col_num)
        cell.value = f'=SUM({col_letter}5:{col_letter}30)'
        cell.font = total_font
        cell.fill = total_fill
        cell.border = total_border
        cell.number_format = '#,##0'

    ws.row_dimensions[31].height = 20

    # Freeze panes at row 5 (freeze header area)
    ws.freeze_panes = 'A5'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
