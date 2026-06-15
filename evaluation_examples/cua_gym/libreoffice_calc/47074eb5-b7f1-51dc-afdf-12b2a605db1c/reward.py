"""
Reward Script: Copy 'Budget Template' sheet three times to create department sheets
Task ID: calc_sht_copy_005
Domain: libreoffice_calc
Scoring:
  - Component 1: All three new department sheets exist (0.4 pts)
  - Component 2: Sheet names match exactly (0.3 pts)
  - Component 3: Sheet order is correct (0.15 pts)
  - Component 4: New sheets contain correct content (SUM formulas intact) (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sht_copy_005'

EXPECTED_SHEETS = ['Budget Template', 'Dept - Engineering', 'Dept - Marketing', 'Dept - Sales']
NEW_SHEETS = ['Dept - Engineering', 'Dept - Marketing', 'Dept - Sales']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    actual_sheets = wb.sheetnames

    # Component 1: All three department sheets exist (0.4 pts)
    # This FAILS on initial (only 1 sheet) and PASSES on golden (4 sheets)
    try:
        sheets_found = [s for s in NEW_SHEETS if s in actual_sheets]
        if len(sheets_found) == 3:
            print(f"PASS: Component 1 — All 3 department sheets exist: {sheets_found} (0.4 pts)")
            total_score += 0.4
        elif len(sheets_found) > 0:
            partial = round(0.4 * len(sheets_found) / 3, 4)
            print(f"PARTIAL: Component 1 — {len(sheets_found)}/3 department sheets found: {sheets_found} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No department sheets found. Present: {actual_sheets}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Exact sheet names match (0.3 pts)
    # The sheets must have the precise names with spaces and dashes as specified
    try:
        all_names_correct = all(s in actual_sheets for s in NEW_SHEETS)
        if all_names_correct:
            print(f"PASS: Component 2 — All sheet names match exactly: {NEW_SHEETS} (0.3 pts)")
            total_score += 0.3
        else:
            missing = [s for s in NEW_SHEETS if s not in actual_sheets]
            print(f"FAIL: Component 2 — Missing sheets with exact names: {missing}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet order is correct (0.15 pts)
    # Expected order: ['Budget Template', 'Dept - Engineering', 'Dept - Marketing', 'Dept - Sales']
    try:
        if actual_sheets == EXPECTED_SHEETS:
            print(f"PASS: Component 3 — Sheet order is correct: {actual_sheets} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Sheet order mismatch. Expected: {EXPECTED_SHEETS}, Found: {actual_sheets}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: New sheets contain correct content - SUM formulas in row 31 intact (0.15 pts)
    # These sheets must be copies of Budget Template, so they must have the same SUM formulas
    # in row 31 (the total row) and the same header in A1.
    # This FAILS on initial (sheets don't exist) and PASSES on golden (copies have correct formulas)
    try:
        content_checks_passed = 0
        content_checks_total = 0
        for sheet_name in NEW_SHEETS:
            if sheet_name not in actual_sheets:
                continue
            ws = wb[sheet_name]
            # Check A1 header value
            a1_val = ws['A1'].value
            # Check SUM formula in B31 (first SUM formula in total row)
            b31_val = ws.cell(row=31, column=2).value
            # Check row 4 header labels
            a4_val = ws.cell(row=4, column=1).value

            content_checks_total += 1
            if (a1_val == 'DEPARTMENT BUDGET 2025' and
                    isinstance(b31_val, str) and 'SUM' in b31_val.upper() and
                    a4_val == 'Line Item'):
                content_checks_passed += 1
            else:
                print(f"  DETAIL: {sheet_name} — A1={repr(a1_val)}, B31={repr(b31_val)}, A4={repr(a4_val)}")

        if content_checks_total == 0:
            print(f"FAIL: Component 4 — No new sheets to verify content")
        elif content_checks_passed == content_checks_total:
            print(f"PASS: Component 4 — All {content_checks_passed} new sheets have correct content (header, formulas) (0.15 pts)")
            total_score += 0.15
        elif content_checks_passed > 0:
            partial = round(0.15 * content_checks_passed / content_checks_total, 4)
            print(f"PARTIAL: Component 4 — {content_checks_passed}/{content_checks_total} sheets have correct content ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No new sheets have correct content")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
