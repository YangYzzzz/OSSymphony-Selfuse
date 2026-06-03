"""
Initial Setup: Create SalesHistory spreadsheet with product sales data (no chart).
Task ID: calc_chart_noncontiguous_range_034
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_noncontiguous_range_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: SalesHistory ---
    ws = wb.active
    ws.title = 'SalesHistory'

    # Headers
    headers = ['Product', '2022 Sales', '2023 Sales', '2024 Sales']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Data rows (exact values from task context)
    data = [
        ['Product Alpha', 12400, 14200, 16800],
        ['Product Beta',   9800, 11200, 13500],
        ['Product Gamma', 18200, 19600, 21400],
        ['Product Delta',  7600,  8900, 10200],
        ['Product Epsilon', 14100, 15800, 18600],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.font = Font(name='Calibri', size=11)
            else:
                cell.number_format = '#,##0'
                cell.font = Font(name='Calibri', size=11)

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # NO charts — task requires creating one

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
