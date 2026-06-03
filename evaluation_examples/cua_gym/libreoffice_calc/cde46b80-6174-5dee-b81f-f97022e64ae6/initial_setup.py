"""
Initial Setup: Change SUM to SUBTOTAL for visible-only rows
Task ID: calc_tbl_023
Domain: libreoffice_calc

Creates a spreadsheet with B1:B24 values totaling 6000.
Rows 5,10,15,20 are hidden with values 500,400,300,200.
B25 has =SUM(B1:B24) which includes hidden rows (shows 6000).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    # Headers
    ws.cell(row=1, column=1, value="Expense Category")
    ws.cell(row=1, column=2, value="Amount")

    # Style headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col in [1, 2]:
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows 2-24 (23 data rows)
    # Hidden rows: 5 (500), 10 (400), 15 (300), 20 (200) => hidden total = 1400
    # Visible total must be 4600, so total = 6000
    # We need 19 visible rows to sum to 4600

    expense_items = [
        ("Office Supplies", 180),        # row 2
        ("Software Licenses", 350),      # row 3
        ("Internet Service", 120),       # row 4
        ("Conference Travel", 500),      # row 5 (HIDDEN)
        ("Team Lunches", 210),           # row 6
        ("Marketing Materials", 275),    # row 7
        ("Cloud Hosting", 450),          # row 8
        ("Phone Plans", 160),            # row 9
        ("Training Workshop", 400),      # row 10 (HIDDEN)
        ("Printing Costs", 95),          # row 11
        ("Equipment Repair", 320),       # row 12
        ("Parking Permits", 140),        # row 13
        ("Health Insurance", 580),       # row 14
        ("Consulting Fees", 300),        # row 15 (HIDDEN)
        ("Courier Services", 85),        # row 16
        ("Building Maintenance", 410),   # row 17
        ("Client Entertainment", 190),   # row 18
        ("Janitorial Services", 155),    # row 19
        ("Legal Retainer", 200),         # row 20 (HIDDEN)
        ("Stationery", 70),              # row 21
        ("Security System", 230),        # row 22
        ("Utilities", 345),              # row 23
        ("Miscellaneous", 235),          # row 24
    ]

    for i, (item, amount) in enumerate(expense_items):
        row = i + 2
        ws.cell(row=row, column=1, value=item)
        ws.cell(row=row, column=2, value=amount)

    # Verify totals
    all_vals = [amt for _, amt in expense_items]
    assert sum(all_vals) == 6000, f"Total is {sum(all_vals)}, expected 6000"
    hidden_vals = [expense_items[3][1], expense_items[8][1], expense_items[13][1], expense_items[18][1]]
    assert hidden_vals == [500, 400, 300, 200], f"Hidden values: {hidden_vals}"

    # B25: SUM formula that includes hidden rows
    ws.cell(row=25, column=1, value="Total")
    ws.cell(row=25, column=1).font = Font(bold=True)
    ws.cell(row=25, column=2, value="=SUM(B1:B24)")
    ws.cell(row=25, column=2).font = Font(bold=True)

    # Format Amount column as currency
    for row in range(2, 26):
        ws.cell(row=row, column=2).number_format = '#,##0'

    # Set column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15

    # Hide rows 5, 10, 15, 20
    ws.row_dimensions[5].hidden = True
    ws.row_dimensions[10].hidden = True
    ws.row_dimensions[15].hidden = True
    ws.row_dimensions[20].hidden = True

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
