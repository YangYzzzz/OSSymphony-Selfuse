"""
Initial Setup: Define named ranges and SUMPRODUCT formula for expense tracking
Task ID: calc_nrv_019
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"

    # Headers
    ws.cell(row=1, column=1, value="Cost Center")
    ws.cell(row=1, column=2, value="Description")
    ws.cell(row=1, column=3, value="Amount")
    ws.cell(row=1, column=6, value="Marketing Total")
    # F2 is intentionally left empty (task requires creating the formula)

    # Realistic expense data - 49 rows (rows 2-50)
    cost_centers = ["Marketing", "Engineering", "Sales", "Operations"]
    descriptions = {
        "Marketing": [
            "Social media campaign", "Print advertising", "Trade show booth",
            "Email marketing platform", "Brand photography", "PR agency retainer",
            "Content creation tools", "Market research survey", "Influencer partnership",
            "Billboard rental", "Google Ads spend", "SEO consulting",
            "Promotional merchandise", "Video production",
        ],
        "Engineering": [
            "AWS hosting fees", "GitHub Enterprise license", "CI/CD pipeline tools",
            "Code review platform", "Testing framework license", "Dev laptop procurement",
            "Technical conference tickets", "Stack Overflow Teams", "Docker Enterprise",
            "Security audit", "Cloud monitoring service", "API gateway subscription",
            "Database migration tool", "Load testing service",
        ],
        "Sales": [
            "CRM subscription", "Sales training workshop", "Client dinner",
            "Demo environment hosting", "Travel expenses Q1", "Commission payout",
            "Lead generation tool", "Sales collateral printing", "Conference sponsorship",
            "Account management software", "Proposal automation tool", "Territory mapping tool",
            "Customer gifting", "Sales analytics platform",
        ],
        "Operations": [
            "Office supplies", "Cleaning service contract", "Facility maintenance",
            "Insurance premium", "Utility bills", "Copier lease",
            "Security system upgrade", "HVAC maintenance", "Parking lot repairs",
            "Fire safety inspection", "Ergonomic furniture", "Waste management service",
            "Building permit renewal", "Reception desk supplies",
        ],
    }

    random.seed(42)  # reproducible

    data = []
    for i in range(49):
        cc = cost_centers[i % 4]
        desc_list = descriptions[cc]
        desc = desc_list[i // 4 % len(desc_list)]
        amount = round(random.uniform(150, 12000), 2)
        data.append((cc, desc, amount))

    for r, (cc, desc, amount) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=cc)
        ws.cell(row=r, column=2, value=desc)
        ws.cell(row=r, column=3, value=amount)

    # NO named ranges in initial (task requires creating them)
    # NO formula in F2 (task requires creating it)

    # Column widths for readability
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["F"].width = 18

    # Bold headers
    from openpyxl.styles import Font
    for col in [1, 2, 3, 6]:
        ws.cell(row=1, column=col).font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
