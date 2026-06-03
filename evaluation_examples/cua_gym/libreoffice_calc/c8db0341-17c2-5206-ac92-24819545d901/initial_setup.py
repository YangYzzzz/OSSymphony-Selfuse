"""
Initial Setup: Create a Gantt chart spreadsheet for home construction project
Task ID: calc_grs_037
Domain: libreoffice_calc

Creates the initial spreadsheet with 12 construction tasks, start weeks,
durations, week headers (1-20), and contractor names. Does NOT include
IF formulas in the week grid, conditional formatting, freeze panes, or
dropdown validation - those are what the agent must create.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # --- Headers ---
    headers_abc = ["Task", "Start Week", "Duration (weeks)"]
    for col, h in enumerate(headers_abc, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # Week headers in columns D (4) through W (23) for Week 1-20
    for week_num in range(1, 21):
        col = week_num + 3  # Week 1 -> col 4 (D), Week 20 -> col 23 (W)
        cell = ws.cell(row=1, column=col, value=week_num)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # Contractor column header (column X = 24)
    contractor_col = 24
    cell = ws.cell(row=1, column=contractor_col, value="Contractor")
    cell.font = Font(bold=True, size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # --- Construction task data ---
    # (Task Name, Start Week, Duration, Contractor)
    tasks = [
        ("Foundation",          1, 3, "Martinez Concrete LLC"),
        ("Framing",             3, 4, "Apex Framing Co."),
        ("Roofing",             6, 2, "Summit Roofing Inc."),
        ("Plumbing Rough-in",   7, 2, "Reliable Plumbing Services"),
        ("Electrical Rough-in", 7, 2, "Brightline Electric"),
        ("Insulation",          9, 1, "ComfortSeal Insulation"),
        ("Drywall",            10, 2, "Apex Framing Co."),
        ("Flooring",           12, 2, "Summit Roofing Inc."),
        ("Interior Painting",  13, 2, "ComfortSeal Insulation"),
        ("Cabinetry",          14, 2, "Reliable Plumbing Services"),
        ("Fixtures",           16, 2, "Brightline Electric"),
        ("Final Inspection",   18, 1, "Martinez Concrete LLC"),
    ]

    for r, (task_name, start_week, duration, contractor) in enumerate(tasks, 2):
        ws.cell(row=r, column=1, value=task_name).font = Font(size=11)
        ws.cell(row=r, column=2, value=start_week).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=3, value=duration).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=contractor_col, value=contractor).font = Font(size=10)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16
    for week_num in range(1, 21):
        col_letter = openpyxl.utils.get_column_letter(week_num + 3)
        ws.column_dimensions[col_letter].width = 6
    ws.column_dimensions[openpyxl.utils.get_column_letter(contractor_col)].width = 28

    # --- Row heights ---
    ws.row_dimensions[1].height = 25

    # --- Light grid borders on data area ---
    thin = Side(style="thin", color="B0B0B0")
    for r in range(1, 14):
        for c in range(1, contractor_col + 1):
            ws.cell(row=r, column=c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
