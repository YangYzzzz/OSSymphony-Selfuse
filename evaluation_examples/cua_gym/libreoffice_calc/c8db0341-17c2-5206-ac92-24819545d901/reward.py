"""
Reward Script: Gantt chart for home construction project in LibreOffice Calc
Task ID: calc_grs_037
Domain: libreoffice_calc
Scoring:
  Component 1: IF formulas in week columns D2:W13 (0.35 points)
  Component 2: Conditional formatting on D2:W13 with dark blue fill (0.25 points)
  Component 3: Freeze panes at D2 (freezes cols A-C) (0.20 points)
  Component 4: Data validation dropdown on contractor column X2:X13 with 6 options (0.20 points)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_037'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition gate: must have 12 task rows (rows 2-13) and week headers
    try:
        if ws.max_row < 13 or ws.max_column < 23:
            print(f"FAIL: Precondition — expected at least 13 rows and 23 cols, "
                  f"found {ws.max_row} rows, {ws.max_column} cols")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"ERROR: Precondition check — {e}")
        print("REWARD: 0.0")
        return 0.0

    # =========================================================================
    # Component 1: IF formulas in week columns D2:W13 (0.35 points)
    # The golden file has IF formulas; the initial file has None in these cells.
    # We check that a significant portion of D2:W13 contain IF formulas.
    # =========================================================================
    try:
        formula_count = 0
        total_cells = 0
        sample_formula = None
        for row in range(2, 14):  # rows 2-13
            for col in range(4, 24):  # cols D(4) through W(23)
                total_cells += 1
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, str) and '=IF(' in val.upper():
                    formula_count += 1
                    if sample_formula is None:
                        sample_formula = val

        # Need at least 80% of cells to have IF formulas
        formula_ratio = formula_count / total_cells if total_cells > 0 else 0
        if formula_ratio >= 0.8:
            print(f"PASS: Component 1 — IF formulas found in {formula_count}/{total_cells} "
                  f"cells ({formula_ratio:.0%}). Sample: {sample_formula} (0.35 pts)")
            total_score += 0.35
        elif formula_ratio >= 0.5:
            partial = 0.35 * (formula_ratio / 0.8)
            if partial > 0:
                print(f"PARTIAL: Component 1 — IF formulas in {formula_count}/{total_cells} "
                      f"cells ({formula_ratio:.0%}), awarding {partial:.2f} pts")
                total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected IF formulas in D2:W13, "
                  f"found formulas in only {formula_count}/{total_cells} cells ({formula_ratio:.0%})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Conditional formatting on D2:W13 with dark blue fill (0.25 points)
    # Initial file has NO conditional formatting. Golden has it.
    # =========================================================================
    try:
        cf_rule_count = 0
        dark_blue_count = 0
        cf_rules = list(ws.conditional_formatting)

        for cf in cf_rules:
            for rule in cf.rules:
                cf_rule_count += 1
                # Check if the fill color is dark blue (003366 or similar dark blue)
                if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                    fill_color = None
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb
                    except Exception:
                        pass
                    if fill_color:
                        # Accept various dark blue shades
                        rgb_hex = fill_color[-6:]  # last 6 chars = RGB
                        r_val = int(rgb_hex[0:2], 16)
                        g_val = int(rgb_hex[2:4], 16)
                        b_val = int(rgb_hex[4:6], 16)
                        # Dark blue: low R, low G, moderate-to-high B (e.g. 003366)
                        if r_val <= 80 and g_val <= 80 and b_val >= 50:
                            dark_blue_count += 1

        if cf_rule_count > 0 and dark_blue_count > 0:
            print(f"PASS: Component 2 — Conditional formatting with dark blue fill detected (0.25 pts)")
            total_score += 0.25
        elif cf_rule_count > 0:
            print(f"PARTIAL: Component 2 — Conditional formatting found but color not dark blue (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — No conditional formatting found on Gantt area")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Freeze panes at D2 (freezes columns A-C and row 1) (0.20 points)
    # Initial file has freeze_panes = None. Golden has D2.
    # =========================================================================
    try:
        freeze = ws.freeze_panes
        if freeze is not None:
            freeze_str = str(freeze).upper()
            if freeze_str == 'D2':
                print(f"PASS: Component 3 — Freeze panes at D2 (columns A-C frozen) (0.20 pts)")
                total_score += 0.20
            else:
                # Partial credit if frozen at some position that at least freezes col A-C
                # D2 means col D row 2 — cols A,B,C frozen, row 1 frozen
                # Accept D1 (only cols frozen) or other D-column freeze
                if freeze_str.startswith('D'):
                    print(f"PARTIAL: Component 3 — Freeze panes at {freeze} (expected D2) (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 3 — Freeze panes at {freeze}, expected D2")
        else:
            print(f"FAIL: Component 3 — No freeze panes set (expected D2)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Data validation dropdown on contractor column with 6 options (0.20 points)
    # Initial file has NO data validation. Golden has list validation on X2:X13.
    # =========================================================================
    try:
        list_dv_count = 0
        option_count = 0

        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list':
                list_dv_count += 1
                # Check that formula1 has multiple options (at least 4 contractor names)
                if dv.formula1:
                    options = str(dv.formula1).strip('"').split(',')
                    option_count = max(option_count, len(options))
                    if len(options) >= 4:
                        print(f"  Data validation options ({len(options)}): {options[:3]}...")

        if list_dv_count > 0 and option_count >= 4:
            print(f"PASS: Component 4 — Data validation dropdown with contractor options (0.20 pts)")
            total_score += 0.20
        elif list_dv_count > 0:
            print(f"PARTIAL: Component 4 — Data validation found but incomplete (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — No list data validation found for contractor dropdown")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state()

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
