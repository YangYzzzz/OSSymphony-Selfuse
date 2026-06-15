"""
Initial Setup: Apply conditional formatting for strikethrough + gray fill on 'Done' rows
Task ID: calc_gg2_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Checklist'

    # Headers
    headers = ['Task Number', 'Task Name', 'Assigned To', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic task data - 50 rows (rows 2-51)
    task_names = [
        'Review Q1 financial report', 'Update client onboarding docs',
        'Migrate database to new server', 'Design landing page mockup',
        'Fix authentication bug in API', 'Prepare quarterly presentation',
        'Conduct user research interviews', 'Optimize search algorithm',
        'Draft partnership agreement', 'Set up CI/CD pipeline',
        'Analyze customer churn data', 'Create employee handbook update',
        'Implement payment gateway', 'Review security audit findings',
        'Plan team offsite agenda', 'Refactor notification service',
        'Write API documentation', 'Test mobile app release candidate',
        'Configure monitoring alerts', 'Design email templates',
        'Update privacy policy', 'Build analytics dashboard',
        'Organize vendor evaluation', 'Resolve server timeout issues',
        'Create training materials', 'Implement SSO integration',
        'Review pull requests backlog', 'Set up staging environment',
        'Draft marketing campaign brief', 'Optimize image loading performance',
        'Conduct accessibility audit', 'Migrate email service provider',
        'Develop feature flag system', 'Create data backup strategy',
        'Review compliance requirements', 'Build customer feedback portal',
        'Set up load testing framework', 'Design onboarding flow',
        'Implement rate limiting', 'Write unit tests for checkout',
        'Create inventory tracking system', 'Audit third-party dependencies',
        'Develop API versioning strategy', 'Configure CDN for assets',
        'Plan product roadmap review', 'Build automated report generator',
        'Implement webhook handler', 'Review database indexing',
        'Set up error tracking service', 'Prepare board presentation',
    ]

    assignees = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'David Kim',
        'Emily Rodriguez', 'James Wilson', 'Aisha Mohammed', 'Carlos Garcia',
        'Rachel Thompson', 'Wei Zhang', 'Olivia Brown', 'Daniel Lee',
        'Fatima Ali', 'Michael Scott', 'Sofia Hernandez',
    ]

    # Status distribution: ~12 Done, ~18 In Progress, ~8 Blocked, ~12 empty
    statuses = [
        'Done', 'In Progress', '', 'Done', 'Blocked',
        'In Progress', 'Done', '', 'In Progress', 'Done',
        'Blocked', 'In Progress', 'Done', 'In Progress', '',
        'Done', 'In Progress', 'Blocked', '', 'In Progress',
        'Done', 'In Progress', '', 'Blocked', 'Done',
        'In Progress', 'Done', 'In Progress', '', 'Blocked',
        'In Progress', 'Done', '', 'In Progress', 'Done',
        'Blocked', 'In Progress', '', 'Done', 'In Progress',
        'Blocked', '', 'In Progress', 'Done', '',
        'In Progress', 'Blocked', 'Done', 'In Progress', '',
    ]

    for i in range(50):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)  # Task Number
        ws.cell(row=row, column=2, value=task_names[i])
        ws.cell(row=row, column=3, value=assignees[i % len(assignees)])
        ws.cell(row=row, column=4, value=statuses[i])

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
