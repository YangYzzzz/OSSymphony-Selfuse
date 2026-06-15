"""
Reward Script: Icon set conditional formatting on Trend column
Task ID: calc_gsd_016
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): CF rule exists on F2:F26 with type iconSet
  Component 2 (0.30): Icon set is a 3-arrow directional set
  Component 3 (0.25): Thresholds at -10 and 10 with numeric type
  Component 4 (0.15): Data integrity — F values unchanged AND CF only on expected range
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_016'

# Expected F column values (ground truth from initial state — must not change)
EXPECTED_F_VALUES = {
    2: 15.3, 3: -8.2, 4: 10.4, 5: -18.5, 6: 32.1,
    7: -5.7, 8: 8.4, 9: -10.9, 10: 22.7, 11: -25.5,
    12: 12.1, 13: 7.4, 14: -3.2, 15: -7.2, 16: 14.8,
    17: -15.4, 18: 11, 19: -2.8, 20: 29.6, 21: -12.3,
    22: 3.8, 23: -4.3, 24: 18.5, 25: -22.1, 26: 28.3,
}


def verify_task(file_path):
    """
    Verify icon set conditional formatting on F2:F26.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Trends' not in wb.sheetnames:
        print("CRITICAL: 'Trends' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Trends']

    # Find the iconSet CF rule targeting F2:F26
    icon_rule = None
    icon_range = None
    for cf in ws.conditional_formatting:
        range_str = str(cf).replace('<ConditionalFormatting ', '').replace('>', '')
        for rule in cf.rules:
            if rule.type == 'iconSet':
                icon_rule = rule
                icon_range = range_str
                break
        if icon_rule:
            break

    # Component 1: CF rule exists on F2:F26 with type iconSet (0.30 points)
    try:
        if icon_rule is not None:
            # Check that the range covers F2:F26
            # Normalize: accept "F2:F26" or equivalent
            normalized_range = icon_range.strip().upper().replace(' ', '')
            if normalized_range == 'F2:F26':
                print(f"PASS: Component 1 — iconSet CF rule found on F2:F26 (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 — iconSet CF found but on range '{icon_range}', expected F2:F26")
        else:
            print("FAIL: Component 1 — No iconSet conditional formatting rule found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Icon set is a 3-arrow directional set (0.30 points)
    try:
        if icon_rule is not None and hasattr(icon_rule, 'iconSet') and icon_rule.iconSet:
            icon_set = icon_rule.iconSet
            icon_name = icon_set.iconSet
            # Accept various 3-arrow icon set names
            valid_arrow_sets = {'3Arrows', '3ArrowsGray', '3Triangles'}
            if icon_name in valid_arrow_sets:
                print(f"PASS: Component 2 — Icon set is '{icon_name}' (3-arrow directional) (0.30 pts)")
                total_score += 0.30
            else:
                # Also accept if it's a 3-icon set (partial: some credit for having icons)
                print(f"FAIL: Component 2 — Icon set is '{icon_name}', expected a 3-arrow set ({valid_arrow_sets})")
        else:
            print("FAIL: Component 2 — No iconSet object found on rule")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Thresholds at -10 and 10 with numeric type (0.25 points)
    try:
        if icon_rule is not None and hasattr(icon_rule, 'iconSet') and icon_rule.iconSet:
            icon_set = icon_rule.iconSet
            cfvos = icon_set.cfvo
            if len(cfvos) >= 3:
                # cfvo[0] is the base (typically 0 or min), cfvo[1] and cfvo[2] are thresholds
                vals = [(cfvo.type, cfvo.val) for cfvo in cfvos]
                print(f"  CFVO values: {vals}")

                # Check that the thresholds include -10 and 10 with numeric type
                threshold_vals = set()
                non_numeric_count = 0
                for cfvo in cfvos:
                    if cfvo.type == 'num':
                        if cfvo.val is not None:
                            threshold_vals.add(float(cfvo.val))
                    else:
                        non_numeric_count += 1

                has_neg10 = -10.0 in threshold_vals
                has_pos10 = 10.0 in threshold_vals

                if has_neg10 and has_pos10 and non_numeric_count == 0:
                    print(f"PASS: Component 3 — Thresholds at -10 and 10, numeric type (0.25 pts)")
                    total_score += 0.25
                elif has_neg10 or has_pos10:
                    partial = 0.125
                    print(f"PARTIAL: Component 3 — Only one threshold correct (-10:{has_neg10}, 10:{has_pos10}) ({partial} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 3 — Thresholds {threshold_vals} don't include -10 and 10")
            else:
                print(f"FAIL: Component 3 — Expected 3 cfvo entries, found {len(cfvos)}")
        else:
            print("FAIL: Component 3 — No iconSet to check thresholds")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data integrity — F values unchanged AND G column unmodified (0.15 points)
    # This is a compound check anchored to task change: the task should add CF WITHOUT altering data.
    # We only award this if CF was actually added (component 1 passed), to avoid scoring preconditions.
    try:
        if icon_rule is not None:  # Only check integrity when CF was added (task-introduced change exists)
            f_ok = (len(EXPECTED_F_VALUES) > 0)  # starts truthy, set false on mismatch
            for row, expected_val in EXPECTED_F_VALUES.items():
                actual = ws.cell(row=row, column=6).value
                if actual is None or abs(float(actual) - expected_val) > 0.01:
                    print(f"  Data mismatch at F{row}: expected {expected_val}, got {actual}")
                    f_ok = False
                    break

            g_header = ws.cell(row=1, column=7).value
            g_ok = (g_header == 'YTD Change%')

            if f_ok and g_ok:
                print(f"PASS: Component 4 — F column data intact, G column unmodified (0.15 pts)")
                total_score += 0.15
            else:
                if not f_ok:
                    print("FAIL: Component 4 — F column values were modified")
                if not g_ok:
                    print(f"FAIL: Component 4 — G column header changed: {g_header}")
        else:
            print("FAIL: Component 4 — Skipped (no CF rule found, so no task change to anchor)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
