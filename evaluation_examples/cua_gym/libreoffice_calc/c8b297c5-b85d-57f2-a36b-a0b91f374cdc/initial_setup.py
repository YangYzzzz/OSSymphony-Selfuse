"""
Initial Setup: Sales dashboard with MoM Change% data for icon set conditional formatting task
Task ID: calc_gsd_016
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trends'

    # Headers
    headers = ['Month', 'Product', 'Region', 'Units Sold', 'Revenue', 'MoM Change%', 'YTD Change%']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 25 rows of realistic monthly sales data (rows 2-26)
    data = [
        ['Jan 2024', 'Widget Pro', 'North America', 1250, 156250.00, 15.3, 15.3],
        ['Jan 2024', 'DataSync Hub', 'Europe', 890, 133500.00, -8.2, -8.2],
        ['Feb 2024', 'Widget Pro', 'North America', 1380, 172500.00, 10.4, 12.8],
        ['Feb 2024', 'DataSync Hub', 'Asia Pacific', 720, 108000.00, -18.5, -13.1],
        ['Mar 2024', 'CloudVault', 'Europe', 2100, 315000.00, 32.1, 22.5],
        ['Mar 2024', 'Widget Pro', 'Asia Pacific', 950, 118750.00, -5.7, 3.4],
        ['Apr 2024', 'DataSync Hub', 'North America', 1100, 165000.00, 8.4, -2.1],
        ['Apr 2024', 'CloudVault', 'Europe', 1870, 280500.00, -10.9, 7.3],
        ['May 2024', 'Widget Pro', 'Europe', 1450, 181250.00, 22.7, 14.6],
        ['May 2024', 'DataSync Hub', 'Asia Pacific', 680, 102000.00, -25.5, -15.8],
        ['Jun 2024', 'CloudVault', 'North America', 2350, 352500.00, 12.1, 18.2],
        ['Jun 2024', 'Widget Pro', 'Asia Pacific', 1020, 127500.00, 7.4, 9.5],
        ['Jul 2024', 'DataSync Hub', 'Europe', 960, 144000.00, -3.2, -6.4],
        ['Jul 2024', 'CloudVault', 'North America', 2180, 327000.00, -7.2, 11.6],
        ['Aug 2024', 'Widget Pro', 'North America', 1560, 195000.00, 14.8, 13.2],
        ['Aug 2024', 'DataSync Hub', 'Asia Pacific', 810, 121500.00, -15.4, -9.7],
        ['Sep 2024', 'CloudVault', 'Europe', 2420, 363000.00, 11.0, 15.8],
        ['Sep 2024', 'Widget Pro', 'Europe', 1290, 161250.00, -2.8, 8.4],
        ['Oct 2024', 'DataSync Hub', 'North America', 1050, 157500.00, 29.6, 2.1],
        ['Oct 2024', 'CloudVault', 'Asia Pacific', 1780, 267000.00, -12.3, 6.9],
        ['Nov 2024', 'Widget Pro', 'North America', 1620, 202500.00, 3.8, 10.1],
        ['Nov 2024', 'DataSync Hub', 'Europe', 920, 138000.00, -4.3, -1.8],
        ['Dec 2024', 'CloudVault', 'North America', 2580, 387000.00, 18.5, 16.4],
        ['Dec 2024', 'Widget Pro', 'Asia Pacific', 1150, 143750.00, -22.1, 4.7],
        ['Dec 2024', 'DataSync Hub', 'North America', 1180, 177000.00, 28.3, 5.2],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 5:  # Revenue
                cell.number_format = '$#,##0.00'
            elif c in (6, 7):  # Percentage columns
                cell.number_format = '0.0'
            elif c == 4:  # Units Sold
                cell.number_format = '#,##0'

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15

    # Freeze header row
    ws.freeze_panes = 'A2'

    # NO conditional formatting - that's the task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
