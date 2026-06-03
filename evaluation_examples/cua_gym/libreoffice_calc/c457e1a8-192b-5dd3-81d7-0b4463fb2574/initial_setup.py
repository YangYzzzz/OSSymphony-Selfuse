"""
Initial Setup: Apply conditional formatting rules to Performance Score column
Task ID: calc_ggf_026
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# Realistic employee names
FIRST_NAMES = [
    "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei", "Carlos",
    "Aisha", "Robert", "Yuki", "Thomas", "Fatima", "Michael", "Sophia",
    "Daniel", "Olivia", "Raj", "Emma", "Alexander", "Lina", "Kevin",
    "Natasha", "Brian", "Zara", "William", "Hannah", "Omar", "Jessica",
    "Nathan", "Amara", "Patrick", "Chloe", "Ahmed", "Victoria", "Lucas",
    "Isabella", "Dmitri", "Grace", "Anthony", "Leila", "Eric", "Mia",
    "Hassan", "Rachel", "Timothy", "Aurora", "Samuel", "Nina", "Victor",
    "Clara", "George", "Lily", "Felix", "Diana", "Hugo", "Aria", "Sanjay",
    "Camille", "Adrian"
]

LAST_NAMES = [
    "Chen", "Johnson", "Petrova", "Williams", "Sharma", "Kim", "Garcia",
    "Nakamura", "Okafor", "Martinez", "Andersen", "Brown", "Taylor",
    "Muller", "Davis", "Patel", "Wilson", "Yamamoto", "Moore", "Lee",
    "Thompson", "Robinson", "Clark", "Rodriguez", "Lewis", "Walker",
    "Hall", "Allen", "Young", "King", "Wright", "Lopez", "Hill", "Scott",
    "Green", "Adams", "Baker", "Gonzalez", "Nelson", "Carter", "Mitchell",
    "Perez", "Roberts", "Turner", "Phillips", "Campbell", "Parker",
    "Evans", "Edwards", "Collins", "Stewart", "Sanchez", "Morris",
    "Rogers", "Reed", "Cook", "Morgan", "Bell", "Murphy", "Bailey"
]


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Metrics"

    # Headers
    headers = ["Employee ID", "Name", "Performance Score"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Generate 120 rows of data
    used_names = set()
    for i in range(120):
        row = i + 2
        emp_id = f"EMP-{1001 + i:04d}"

        # Generate unique names
        while True:
            fn = random.choice(FIRST_NAMES)
            ln = random.choice(LAST_NAMES)
            full_name = f"{fn} {ln}"
            if full_name not in used_names:
                used_names.add(full_name)
                break

        # Performance Score: mostly 1-200, but about 15% are 0 (no data recorded)
        if random.random() < 0.15:
            score = 0
        else:
            score = random.randint(12, 198)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=full_name)
        ws.cell(row=row, column=3, value=score)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
