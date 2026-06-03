"""
Initial Setup: Reorder month sheets, add Q1 Summary with headers and purple tab
Task ID: calc_ps_095
Domain: libreoffice_calc

Initial state: Sheets in wrong order: 'March', 'January', 'February'.
Each has budget data with headers Category, Budget, Actual.
No Q1 Summary sheet exists.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: March (deliberately first to create wrong order) ---
    ws_mar = wb.active
    ws_mar.title = 'March'
    headers = ['Category', 'Budget', 'Actual']
    for col, h in enumerate(headers, 1):
        cell = ws_mar.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    march_data = [
        ['Office Supplies', 1200, 1087.50],
        ['Travel', 8500, 9230.00],
        ['Software Licenses', 3400, 3400.00],
        ['Marketing Campaigns', 15000, 14250.75],
        ['Client Entertainment', 2000, 2415.00],
        ['Professional Development', 4500, 3890.00],
        ['Utilities', 1800, 1765.20],
        ['Equipment Maintenance', 3200, 2950.00],
        ['Contractor Fees', 12000, 11800.00],
        ['Miscellaneous', 1500, 1620.35],
        ['Insurance Premiums', 2800, 2800.00],
        ['Shipping & Logistics', 5600, 5430.10],
    ]
    for r, row_data in enumerate(march_data, 2):
        ws_mar.cell(row=r, column=1, value=row_data[0])
        ws_mar.cell(row=r, column=2, value=row_data[1])
        ws_mar.cell(row=r, column=3, value=row_data[2])

    # Set column widths for readability
    ws_mar.column_dimensions['A'].width = 25
    ws_mar.column_dimensions['B'].width = 14
    ws_mar.column_dimensions['C'].width = 14

    # --- Sheet 2: January ---
    ws_jan = wb.create_sheet('January')
    for col, h in enumerate(headers, 1):
        cell = ws_jan.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    january_data = [
        ['Office Supplies', 1200, 1150.25],
        ['Travel', 8500, 7820.00],
        ['Software Licenses', 3400, 3400.00],
        ['Marketing Campaigns', 15000, 13500.00],
        ['Client Entertainment', 2000, 1875.50],
        ['Professional Development', 4500, 4200.00],
        ['Utilities', 1800, 1920.75],
        ['Equipment Maintenance', 3200, 3100.00],
        ['Contractor Fees', 12000, 12000.00],
        ['Miscellaneous', 1500, 1345.80],
        ['Insurance Premiums', 2800, 2800.00],
        ['Shipping & Logistics', 5600, 5780.45],
    ]
    for r, row_data in enumerate(january_data, 2):
        ws_jan.cell(row=r, column=1, value=row_data[0])
        ws_jan.cell(row=r, column=2, value=row_data[1])
        ws_jan.cell(row=r, column=3, value=row_data[2])

    ws_jan.column_dimensions['A'].width = 25
    ws_jan.column_dimensions['B'].width = 14
    ws_jan.column_dimensions['C'].width = 14

    # --- Sheet 3: February ---
    ws_feb = wb.create_sheet('February')
    for col, h in enumerate(headers, 1):
        cell = ws_feb.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    february_data = [
        ['Office Supplies', 1200, 1098.00],
        ['Travel', 8500, 8150.30],
        ['Software Licenses', 3400, 3400.00],
        ['Marketing Campaigns', 15000, 14800.00],
        ['Client Entertainment', 2000, 2100.00],
        ['Professional Development', 4500, 4050.75],
        ['Utilities', 1800, 1850.50],
        ['Equipment Maintenance', 3200, 3050.00],
        ['Contractor Fees', 12000, 11500.00],
        ['Miscellaneous', 1500, 1480.90],
        ['Insurance Premiums', 2800, 2800.00],
        ['Shipping & Logistics', 5600, 5540.20],
    ]
    for r, row_data in enumerate(february_data, 2):
        ws_feb.cell(row=r, column=1, value=row_data[0])
        ws_feb.cell(row=r, column=2, value=row_data[1])
        ws_feb.cell(row=r, column=3, value=row_data[2])

    ws_feb.column_dimensions['A'].width = 25
    ws_feb.column_dimensions['B'].width = 14
    ws_feb.column_dimensions['C'].width = 14

    # Sheet order is: March, January, February (wrong order - task requires reordering)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet order: {wb.sheetnames}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
