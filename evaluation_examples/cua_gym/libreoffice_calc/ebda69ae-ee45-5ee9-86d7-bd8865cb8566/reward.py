"""
Reward Script: Reorder month sheets to calendar order, add Q1 Summary sheet with headers and purple tab
Task ID: calc_ps_095
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Sheets in correct calendar order (January, February, March first three)
  Component 2 (0.20): 'Q1 Summary' sheet exists as the 4th sheet
  Component 3 (0.25): 'Q1 Summary' row 1 has correct headers from January
  Component 4 (0.25): 'Q1 Summary' tab color is purple
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_095'


def persist_app_state(domain: str):
    """Attempt to save any unsaved GUI state via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Sheet names found: {sheet_names}")

    # Component 1: First three sheets are in calendar order (0.30 points)
    # Initial has ['March', 'January', 'February'] — this check FAILS on initial
    # Golden has ['January', 'February', 'March', 'Q1 Summary'] — PASSES
    try:
        expected_order = ['January', 'February', 'March']
        if len(sheet_names) >= 3 and sheet_names[:3] == expected_order:
            print(f"PASS: Component 1 — Sheets in calendar order: {sheet_names[:3]} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected first 3 sheets {expected_order}, found {sheet_names[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Q1 Summary' sheet exists as the 4th sheet (0.20 points)
    # Initial has no 'Q1 Summary' — FAILS on initial
    # Golden has it as 4th sheet — PASSES
    try:
        if 'Q1 Summary' in sheet_names:
            q1_idx = sheet_names.index('Q1 Summary')
            if q1_idx == 3:
                print(f"PASS: Component 2 — 'Q1 Summary' exists at position 4 (index 3) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — 'Q1 Summary' exists but at position {q1_idx+1}, expected position 4")
        else:
            print(f"FAIL: Component 2 — 'Q1 Summary' sheet not found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 'Q1 Summary' row 1 has correct headers (0.25 points)
    # Expected: A1='Category', B1='Budget', C1='Actual' (copied from January)
    # Initial has no Q1 Summary — FAILS
    try:
        if 'Q1 Summary' in sheet_names:
            ws = wb['Q1 Summary']
            a1 = ws.cell(row=1, column=1).value
            b1 = ws.cell(row=1, column=2).value
            c1 = ws.cell(row=1, column=3).value
            expected_headers = ['Category', 'Budget', 'Actual']
            actual_headers = [a1, b1, c1]
            if actual_headers == expected_headers:
                print(f"PASS: Component 3 — Q1 Summary headers correct: {actual_headers} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Expected headers {expected_headers}, found {actual_headers}")
        else:
            print(f"FAIL: Component 3 — 'Q1 Summary' sheet not found, cannot check headers")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: 'Q1 Summary' tab color is purple (0.25 points)
    # Purple is typically RGB 800080 or similar purple hues
    # Initial has no Q1 Summary — FAILS
    try:
        if 'Q1 Summary' in sheet_names:
            ws = wb['Q1 Summary']
            tab_color = ws.sheet_properties.tabColor
            if tab_color is not None:
                color_rgb = tab_color.rgb if hasattr(tab_color, 'rgb') and tab_color.rgb else None
                print(f"INFO: Q1 Summary tab color rgb: {color_rgb}")
                if color_rgb is not None:
                    # Extract the RGB portion (last 6 chars of ARGB)
                    rgb_hex = str(color_rgb)[-6:].upper()
                    # Check for purple-ish colors:
                    # Classic purple: 800080
                    # Also accept: 7030A0 (Excel purple), 9B30FF, 8B008B, etc.
                    r_val = int(rgb_hex[0:2], 16)
                    g_val = int(rgb_hex[2:4], 16)
                    b_val = int(rgb_hex[4:6], 16)
                    # Purple: red > 0, green is low, blue > 0, and red+blue much greater than green
                    is_purple = (r_val >= 64 and b_val >= 64 and g_val < 100
                                 and (r_val + b_val) > 2 * g_val)
                    if is_purple:
                        print(f"PASS: Component 4 — Q1 Summary tab color is purple (#{rgb_hex}) (0.25 pts)")
                        total_score += 0.25
                    else:
                        print(f"FAIL: Component 4 — Tab color #{rgb_hex} (R={r_val},G={g_val},B={b_val}) does not appear purple")
                else:
                    print(f"FAIL: Component 4 — Tab color exists but rgb is None (theme color?)")
            else:
                print(f"FAIL: Component 4 — Q1 Summary has no tab color set")
        else:
            print(f"FAIL: Component 4 — 'Q1 Summary' sheet not found, cannot check tab color")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
