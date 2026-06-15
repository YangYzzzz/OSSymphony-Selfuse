"""
Initial Setup: Dropdown validation for Orders sheet using ProductList range
Task ID: calc_dop_validate_range_020
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_validate_range_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Orders ---
    ws_orders = wb.active
    ws_orders.title = 'Orders'

    # Headers
    headers = ['Order ID', 'Product Name', 'Quantity', 'Unit Price', 'Total']
    for col, h in enumerate(headers, 1):
        cell = ws_orders.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Partial data rows (rows 2-20 have some data, rest empty)
    sample_orders = [
        ['ORD-10001', None, 2, 899.99, None],
        ['ORD-10002', None, 5, 29.99, None],
        ['ORD-10003', None, 1, 49.99, None],
        ['ORD-10004', None, 3, 149.99, None],
        ['ORD-10005', None, 10, 12.50, None],
        ['ORD-10006', None, 2, 359.00, None],
        ['ORD-10007', None, 4, 89.99, None],
        ['ORD-10008', None, 1, 199.99, None],
        ['ORD-10009', None, 6, 24.99, None],
        ['ORD-10010', None, 2, 75.00, None],
        ['ORD-10011', None, 1, 599.99, None],
        ['ORD-10012', None, 3, 179.99, None],
        ['ORD-10013', None, 8, 18.99, None],
        ['ORD-10014', None, 1, 849.00, None],
        ['ORD-10015', None, 2, 65.00, None],
        ['ORD-10016', None, 1, 249.99, None],
        ['ORD-10017', None, 4, 34.99, None],
        ['ORD-10018', None, 5, 22.00, None],
        ['ORD-10019', None, 2, 39.99, None],
    ]
    # Note: Product Name (column B) is intentionally left empty (None) — no validation yet
    for r, row_data in enumerate(sample_orders, 2):
        for c, val in enumerate(row_data, 1):
            ws_orders.cell(row=r, column=c, value=val)

    # Set column widths
    ws_orders.column_dimensions['A'].width = 14
    ws_orders.column_dimensions['B'].width = 22
    ws_orders.column_dimensions['C'].width = 12
    ws_orders.column_dimensions['D'].width = 12
    ws_orders.column_dimensions['E'].width = 12

    # --- Sheet 2: ProductList ---
    ws_products = wb.create_sheet('ProductList')

    # Header
    ws_products['A1'] = 'Product'
    ws_products['A1'].font = Font(bold=True)

    # 19 product names in A2:A20
    product_names = [
        'Laptop Pro 15',
        'Wireless Mouse',
        'USB-C Hub',
        'Mechanical Keyboard',
        '27-Inch Monitor',
        'Webcam HD',
        'External SSD 1TB',
        'Laptop Stand',
        'Desk Lamp LED',
        'Cable Management Kit',
        'Drawing Tablet',
        'Stream Deck',
        'Blue Light Glasses',
        'Ergonomic Chair',
        'Footrest',
        'Monitor Arm',
        'Headphone Stand',
        'Desktop Fan',
        'Cable Tester',
    ]
    for row_idx, name in enumerate(product_names, 2):
        ws_products.cell(row=row_idx, column=1, value=name)

    ws_products.column_dimensions['A'].width = 25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Orders sheet: headers + 19 sample rows (column B empty, no validation)')
    print(f'  ProductList sheet: header + 19 product names in A2:A20')


create_initial()
