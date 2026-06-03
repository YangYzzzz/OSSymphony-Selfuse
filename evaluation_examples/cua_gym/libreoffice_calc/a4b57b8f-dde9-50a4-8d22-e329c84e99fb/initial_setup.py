"""
Initial Setup: Merge cells A1:F1, center text, type title in 18pt bold Arial
Task ID: calc_gg3_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Report ---
    ws = wb.active
    ws.title = 'Report'

    # Row 1: leftover single characters from an earlier draft (NOT merged, NOT styled)
    leftover_chars = ['Q', '3', ' ', 'S', 'a', 'l']
    for col, ch in enumerate(leftover_chars, 1):
        ws.cell(row=1, column=col, value=ch)

    # Row 2: Headers for data table
    headers = ['Region', 'Sales Rep', 'Q3 Revenue', 'Units Sold', 'Target', 'Achievement %']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
        ws.cell(row=2, column=col).font = Font(bold=True)

    # Rows 3-14: Realistic sales data
    data = [
        ['Northeast', 'Sarah Chen', 87450, 412, 80000, 109.3],
        ['Northeast', 'Marcus Johnson', 76230, 358, 75000, 101.6],
        ['Southeast', 'Priya Patel', 92100, 445, 85000, 108.4],
        ['Southeast', 'James Rodriguez', 68900, 324, 70000, 98.4],
        ['Midwest', 'Emily Thompson', 71550, 338, 72000, 99.4],
        ['Midwest', 'David Kim', 83200, 396, 80000, 104.0],
        ['West', 'Rachel Martinez', 95800, 463, 90000, 106.4],
        ['West', 'Tyler Washington', 78400, 371, 78000, 100.5],
        ['Central', 'Aisha Okafor', 64300, 305, 65000, 98.9],
        ['Central', 'Brian Nakamura', 89750, 427, 85000, 105.6],
        ['Northwest', 'Olivia Fischer', 73600, 349, 74000, 99.5],
        ['Northwest', 'Carlos Mendez', 81900, 389, 80000, 102.4],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Format currency columns
            if c == 3 or c == 5:
                cell.number_format = '$#,##0'
            elif c == 6:
                cell.number_format = '0.0'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
