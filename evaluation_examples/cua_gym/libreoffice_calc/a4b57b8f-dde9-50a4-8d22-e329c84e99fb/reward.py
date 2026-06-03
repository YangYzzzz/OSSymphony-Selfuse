"""
Reward Script: Merge cells A1:F1, center text, type title in 18pt bold Arial
Task ID: calc_gg3_004
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): A1:F1 merged into single cell
  Component 2 (0.25): A1 contains 'Q3 Regional Sales Report'
  Component 3 (0.25): Font is Arial, 18pt, bold
  Component 4 (0.20): Horizontal and vertical center alignment
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_004'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Report' sheet must exist
    if 'Report' not in wb.sheetnames:
        print(f"CRITICAL: 'Report' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: A1:F1 merged into a single cell (0.30 points)
    try:
        merged_ranges = list(ws.merged_cells.ranges)
        # Check that there is a merged range covering A1:F1
        matching = [mr for mr in merged_ranges
                    if mr.min_row == 1 and mr.max_row == 1 and mr.min_col == 1 and mr.max_col == 6]

        if len(matching) > 0:
            # Also verify B1-F1 are MergedCell objects (not top-left)
            all_merged = all(isinstance(ws.cell(row=1, column=c), MergedCell) for c in range(2, 7))
            if all_merged:
                print(f"PASS: Component 1 — A1:F1 merged correctly (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 1 — Merged range A1:F1 found but B1-F1 not all MergedCell")
        else:
            print(f"FAIL: Component 1 — A1:F1 not merged. Merged ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A1 contains 'Q3 Regional Sales Report' (0.25 points)
    try:
        a1_val = ws['A1'].value
        if a1_val is not None and str(a1_val).strip() == 'Q3 Regional Sales Report':
            print(f"PASS: Component 2 — A1 = 'Q3 Regional Sales Report' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Expected 'Q3 Regional Sales Report', found: {repr(a1_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Font is Arial, 18pt, bold (0.25 points)
    try:
        cell = ws['A1']
        font_name = cell.font.name
        font_size = cell.font.size
        font_bold = cell.font.bold

        checks_passed = 0
        details = []

        # Check font name (Arial or close sans-serif equivalent)
        if font_name and font_name.lower() == 'arial':
            checks_passed += 1
            details.append(f"name=Arial OK")
        else:
            details.append(f"name={font_name} (expected Arial)")

        # Check font size = 18
        if font_size is not None and abs(float(font_size) - 18.0) < 0.5:
            checks_passed += 1
            details.append(f"size=18 OK")
        else:
            details.append(f"size={font_size} (expected 18)")

        # Check bold
        if font_bold:
            checks_passed += 1
            details.append(f"bold=True OK")
        else:
            details.append(f"bold={font_bold} (expected True)")

        if checks_passed == 3:
            print(f"PASS: Component 3 — Font correct: {', '.join(details)} (0.25 pts)")
            total_score += 0.25
        elif checks_passed >= 2:
            # Partial credit: 2 of 3 font properties correct
            if checks_passed == 2:
                total_score += 0.17
                print(f"PARTIAL: Component 3 — 2/3 font checks: {', '.join(details)} (0.17 pts)")
        else:
            print(f"FAIL: Component 3 — Font: {', '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Horizontal center + vertical center alignment (0.20 points)
    try:
        cell = ws['A1']
        h_align = cell.alignment.horizontal
        v_align = cell.alignment.vertical

        h_ok = h_align == 'center'
        v_ok = v_align == 'center'

        if h_ok and v_ok:
            print(f"PASS: Component 4 — Alignment h=center, v=center (0.20 pts)")
            total_score += 0.20
        elif h_ok:
            # Only horizontal center correct
            total_score += 0.10
            print(f"PARTIAL: Component 4 — h=center OK, v={v_align} wrong (0.10 pts)")
        elif v_ok:
            # Only vertical center correct
            total_score += 0.10
            print(f"PARTIAL: Component 4 — h={h_align} wrong, v=center OK (0.10 pts)")
        else:
            print(f"FAIL: Component 4 — h={h_align}, v={v_align} (expected center/center)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
