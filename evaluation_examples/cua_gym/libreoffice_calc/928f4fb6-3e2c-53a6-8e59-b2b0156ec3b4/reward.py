"""
Reward Script: Write LibreOffice Basic macros for report_automation.xlsx
Task ID: osworld_multi_apps_docx_to_calc_013
Domain: libreoffice_calc (multi-app: docx + calc)
Scoring:
  Component 1 (0.30): Sheet1 data sorted by date (ascending) — result of macro 1
  Component 2 (0.40): Sheet2 has monthly summary table (Month, Total Amount, Transaction Count)
                       with correct data for all 5 months — result of macro 2
  Component 3 (0.30): monthly_summary.pdf exists on Desktop — result of macro 3
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_docx_to_calc_013'

XLSX_PATH = f'{WORKDIR}/report_automation.xlsx'
PDF_PATH  = f'{WORKDIR}/Desktop/monthly_summary.pdf'

# Expected monthly summary data (from ground truth analysis)
EXPECTED_SUMMARY = [
    ('2025-01', 3687.89, 6),
    ('2025-02', 6559.99, 5),
    ('2025-03', 2333.15, 5),
    ('2025-04', 3522.7,  5),
    ('2025-05', 5060.0,  4),
]
NUMERIC_TOLERANCE = 0.05  # Allow small rounding differences


def verify_task(xlsx_path, pdf_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: xlsx file must exist
    if not os.path.exists(xlsx_path):
        print(f"CRITICAL: Workbook not found at {xlsx_path}")
        print("REWARD: 0.0")
        return 0.0

    try:
        wb = openpyxl.load_workbook(xlsx_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {xlsx_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Sheet1 data sorted by date ascending (0.30 points)
    # This verifies the result of macro 1 (auto-sort data by date).
    # Initial env: data is in random order → fails.
    # Golden env: data sorted chronologically → passes.
    # -------------------------------------------------------------------------
    try:
        if 'Sheet1' not in wb.sheetnames:
            print("FAIL: Component 1 — 'Sheet1' not found in workbook")
        else:
            ws1 = wb['Sheet1']
            dates = []
            for r in range(2, ws1.max_row + 1):
                val = ws1.cell(r, 1).value
                if val is not None:
                    dates.append(str(val))

            if len(dates) == 0:
                print("FAIL: Component 1 — Sheet1 has no data rows")
            elif dates == sorted(dates):
                print(f"PASS: Component 1 — Sheet1 sorted by date ascending "
                      f"({len(dates)} rows, first={dates[0]}, last={dates[-1]}) (+0.30 pts)")
                total_score += 0.30
            else:
                # Identify first out-of-order pair for feedback
                first_bad = None
                for i in range(len(dates) - 1):
                    if dates[i] > dates[i + 1]:
                        first_bad = (i + 2, dates[i], i + 3, dates[i + 1])
                        break
                print(f"FAIL: Component 1 — Sheet1 NOT sorted by date. "
                      f"First out-of-order: row {first_bad[0]}={first_bad[1]} > "
                      f"row {first_bad[2]}={first_bad[3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Sheet2 has monthly summary table (0.40 points)
    # Awarded in two sub-parts:
    #   2a (0.20): Correct headers in Sheet2 (Month, Total Amount, Transaction Count)
    #   2b (0.20): Correct data values for all 5 months (within tolerance)
    # Initial env: Sheet2 is empty → fails both sub-parts.
    # Golden env: Sheet2 populated correctly → passes both sub-parts.
    # -------------------------------------------------------------------------
    try:
        if 'Sheet2' not in wb.sheetnames:
            print("FAIL: Component 2 — 'Sheet2' not found in workbook")
        else:
            ws2 = wb['Sheet2']

            # Sub-component 2a: Headers check (0.20 points)
            h1 = ws2.cell(1, 1).value
            h2 = ws2.cell(1, 2).value
            h3 = ws2.cell(1, 3).value

            expected_headers = ('Month', 'Total Amount', 'Transaction Count')
            actual_headers = (
                str(h1).strip() if h1 else None,
                str(h2).strip() if h2 else None,
                str(h3).strip() if h3 else None,
            )

            headers_ok = (
                h1 is not None and
                h2 is not None and
                h3 is not None and
                str(h1).strip().lower() == 'month' and
                'amount' in str(h2).lower() and
                'count' in str(h3).lower()
            )

            if headers_ok:
                print(f"PASS: Component 2a — Sheet2 headers correct: "
                      f"{actual_headers} (+0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2a — Sheet2 headers incorrect. "
                      f"Expected {expected_headers}, found {actual_headers}")

            # Sub-component 2b: Data values check (0.20 points)
            data_rows = []
            for r in range(2, ws2.max_row + 1):
                m_val  = ws2.cell(r, 1).value
                a_val  = ws2.cell(r, 2).value
                c_val  = ws2.cell(r, 3).value
                if m_val is not None:
                    data_rows.append((str(m_val).strip(), a_val, c_val))

            if len(data_rows) == 0:
                print("FAIL: Component 2b — Sheet2 has no data rows (no monthly data)")
            elif len(data_rows) < len(EXPECTED_SUMMARY):
                print(f"FAIL: Component 2b — Sheet2 has {len(data_rows)} data rows, "
                      f"expected {len(EXPECTED_SUMMARY)}")
            else:
                mismatches = []
                for exp_month, exp_amount, exp_count in EXPECTED_SUMMARY:
                    # Find matching row
                    matched = None
                    for actual_month, actual_amount, actual_count in data_rows:
                        if actual_month == exp_month:
                            matched = (actual_month, actual_amount, actual_count)
                            break

                    if matched is None:
                        mismatches.append(f"Month {exp_month} not found")
                        continue

                    act_month, act_amount, act_count = matched

                    # Check amount
                    try:
                        if abs(float(act_amount) - exp_amount) > NUMERIC_TOLERANCE:
                            mismatches.append(
                                f"{exp_month} Total Amount: expected {exp_amount}, "
                                f"got {act_amount}"
                            )
                    except (TypeError, ValueError):
                        mismatches.append(
                            f"{exp_month} Total Amount: invalid value {act_amount}"
                        )

                    # Check count
                    try:
                        if int(act_count) != exp_count:
                            mismatches.append(
                                f"{exp_month} Transaction Count: expected {exp_count}, "
                                f"got {act_count}"
                            )
                    except (TypeError, ValueError):
                        mismatches.append(
                            f"{exp_month} Transaction Count: invalid value {act_count}"
                        )

                if not mismatches:
                    print(f"PASS: Component 2b — Sheet2 monthly data correct for all "
                          f"{len(EXPECTED_SUMMARY)} months (+0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 2b — Sheet2 data mismatches: "
                          f"{'; '.join(mismatches)}")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: monthly_summary.pdf exported to Desktop (0.30 points)
    # This verifies the result of macro 3 (export Sheet2 as PDF to desktop).
    # Initial env: no PDF on desktop → fails.
    # Golden env: PDF present and valid on desktop → passes.
    # We verify both existence and that the file is a valid PDF (starts with %PDF).
    # -------------------------------------------------------------------------
    try:
        pdf_exists = os.path.exists(pdf_path)
        pdf_valid = False
        if pdf_exists:
            try:
                with open(pdf_path, 'rb') as f:
                    header = f.read(4)
                    pdf_valid = header == b'%PDF'
            except Exception:
                pdf_valid = False

        if pdf_exists and pdf_valid:
            size_kb = os.path.getsize(pdf_path) / 1024
            print(f"PASS: Component 3 — PDF exported, valid (size={size_kb:.1f} KB) (+0.30 pts)")
            total_score += 0.30  # score increment: PDF exists and is a valid PDF
        elif pdf_exists and not pdf_valid:
            print(f"FAIL: Component 3 — monthly_summary.pdf exists but is not a valid PDF "
                  f"(missing %PDF header)")
        else:
            print(f"FAIL: Component 3 — monthly_summary.pdf not found at {pdf_path}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if __name__ == '__main__':
    verify_task(XLSX_PATH, PDF_PATH)
