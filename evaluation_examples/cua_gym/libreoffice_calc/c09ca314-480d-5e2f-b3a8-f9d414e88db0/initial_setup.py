"""
Initial Setup: Department budget spreadsheet with variance data, no conditional formatting.
Task ID: calc_gsd_019
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget vs Actual"

    # Headers
    headers = ["Department", "Budgeted Amount", "Actual Amount", "Variance $", "Variance %"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_color = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_color
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 20 department budget entries with realistic data
    departments = [
        ("Human Resources", 185000, 192500),
        ("Marketing", 320000, 298750),
        ("Engineering", 540000, 567200),
        ("Sales", 275000, 261300),
        ("Finance", 195000, 195000),
        ("Legal", 210000, 224800),
        ("Customer Support", 165000, 158400),
        ("Research & Development", 480000, 512300),
        ("Operations", 290000, 275600),
        ("IT Infrastructure", 350000, 363400),
        ("Product Management", 230000, 218500),
        ("Quality Assurance", 175000, 182900),
        ("Supply Chain", 245000, 239100),
        ("Public Relations", 128000, 135600),
        ("Training & Development", 95000, 88200),
        ("Facilities Management", 210000, 217800),
        ("Data Analytics", 185000, 176300),
        ("Compliance", 155000, 162400),
        ("Executive Office", 420000, 415000),
        ("Procurement", 198000, 203500),
    ]

    for r, (dept, budgeted, actual) in enumerate(departments, 2):
        variance_dollar = budgeted - actual  # positive = under budget, negative = over budget
        variance_pct = (variance_dollar / budgeted) * 100 if budgeted != 0 else 0

        ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=2, value=budgeted)
        ws.cell(row=r, column=3, value=actual)
        ws.cell(row=r, column=4, value=variance_dollar)
        ws.cell(row=r, column=5, value=round(variance_pct, 2))

        # Number formatting
        ws.cell(row=r, column=2).number_format = '$#,##0.00'
        ws.cell(row=r, column=3).number_format = '$#,##0.00'
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
        ws.cell(row=r, column=5).number_format = '0.00"%"'

    # Column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting - that is the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
