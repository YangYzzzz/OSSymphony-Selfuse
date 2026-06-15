"""
Reward Script: Unlock specific cells for user input on Registration sheet and protect with password
Task ID: calc_ps_022
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Sheet protection is enabled
  Component 2 (0.3): Input cells B3, B5, B7, B9, B11 are unlocked
  Component 3 (0.2): Input cells D3, D5, D7, D9, D11 are unlocked
  Component 4 (0.2): Label cells (A-col and C-col) remain locked
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_022'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Registration' sheet exists
    if 'Registration' not in wb.sheetnames:
        print("FAIL: 'Registration' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Registration']

    # Component 1: Sheet protection is enabled (0.3 points)
    # Initial: protection.sheet == False; Golden: protection.sheet == True
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 — Sheet protection is enabled (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Sheet protection is NOT enabled (protection.sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Input cells B3, B5, B7, B9, B11 are unlocked (0.3 points)
    # Initial: all locked=True; Golden: all locked=False
    try:
        b_cells = ['B3', 'B5', 'B7', 'B9', 'B11']
        unlocked_count = 0
        for coord in b_cells:
            cell = ws[coord]
            if not cell.protection.locked:
                unlocked_count += 1
            else:
                print(f"  INFO: {coord} is still locked (expected unlocked)")

        if unlocked_count == len(b_cells):
            print(f"PASS: Component 2 — All B-column input cells unlocked ({unlocked_count}/{len(b_cells)}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Only {unlocked_count}/{len(b_cells)} B-column input cells unlocked")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Input cells D3, D5, D7, D9, D11 are unlocked (0.2 points)
    # Initial: all locked=True; Golden: all locked=False
    try:
        d_cells = ['D3', 'D5', 'D7', 'D9', 'D11']
        unlocked_count = 0
        for coord in d_cells:
            cell = ws[coord]
            if not cell.protection.locked:
                unlocked_count += 1
            else:
                print(f"  INFO: {coord} is still locked (expected unlocked)")

        if unlocked_count == len(d_cells):
            print(f"PASS: Component 3 — All D-column input cells unlocked ({unlocked_count}/{len(d_cells)}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {unlocked_count}/{len(d_cells)} D-column input cells unlocked")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Label cells remain locked (0.2 points)
    # Both initial and golden have these locked=True, BUT this component is gated
    # on sheet protection being enabled (Component 1). On initial_env, protection is
    # off so locking is irrelevant — we only award points if protection is active AND
    # label cells are locked. This ensures initial_env scores 0 for this component.
    try:
        label_cells = ['A3', 'A5', 'A7', 'A9', 'A11', 'C3', 'C5', 'C7', 'C9', 'C11']
        unlocked_labels = [c for c in label_cells if not ws[c].protection.locked]
        for coord in unlocked_labels:
            print(f"  INFO: {coord} is unlocked (expected locked)")

        # Gate on protection being active — if sheet isn't protected, locked status is meaningless
        if ws.protection.sheet and len(unlocked_labels) == 0:
            print(f"PASS: Component 4 — All label cells remain locked with protection active (0.2 pts)")
            total_score += 0.2
        elif not ws.protection.sheet:
            print(f"FAIL: Component 4 — Sheet protection not active, locked status is irrelevant")
        else:
            print(f"FAIL: Component 4 — {len(unlocked_labels)} label cells are unlocked: {unlocked_labels}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
