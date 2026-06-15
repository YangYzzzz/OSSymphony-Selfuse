"""
Initial Setup: Create registration form template with labels
Task ID: calc_ps_022
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registration"

    # === Header row ===
    ws.merge_cells("A1:D1")
    ws["A1"] = "Employee Registration Form"
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # === Column A labels ===
    labels_a = {
        3: "Name",
        5: "Email",
        7: "Phone",
        9: "Department",
        11: "Start Date",
    }
    for row, label in labels_a.items():
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # === Column C labels ===
    labels_c = {
        3: "Title",
        5: "Manager",
        7: "Location",
        9: "Badge#",
        11: "Notes",
    }
    for row, label in labels_c.items():
        cell = ws.cell(row=row, column=3, value=label)
        cell.font = Font(name="Arial", size=11, bold=True)
        cell.alignment = Alignment(horizontal="right", vertical="center")

    # === Style input cells with light background to indicate editable areas ===
    input_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="999999"),
    )
    input_cells_b = [3, 5, 7, 9, 11]
    for row in input_cells_b:
        cell = ws.cell(row=row, column=2)
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    input_cells_d = [3, 5, 7, 9, 11]
    for row in input_cells_d:
        cell = ws.cell(row=row, column=4)
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center")

    # === Column widths ===
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 25

    # === Row heights for spacing ===
    ws.row_dimensions[1].height = 35
    for r in [3, 5, 7, 9, 11]:
        ws.row_dimensions[r].height = 22
    for r in [2, 4, 6, 8, 10]:
        ws.row_dimensions[r].height = 8  # spacer rows

    # === Footer instruction ===
    ws["A13"] = "Please fill in all fields above and submit to HR."
    ws["A13"].font = Font(name="Arial", size=9, italic=True, color="888888")
    ws.merge_cells("A13:D13")

    # All cells are locked by default in openpyxl.
    # Sheet is NOT protected (task requires the agent to protect it).

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
