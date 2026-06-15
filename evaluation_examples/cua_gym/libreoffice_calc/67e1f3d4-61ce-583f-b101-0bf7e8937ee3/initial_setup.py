"""
Initial Setup: Create a spreadsheet with two separate financial tables (no charts yet)
Task ID: osworld_calc_dual_chart_separate_tables_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_dual_chart_separate_tables_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Data"

    # --- Operating Expenses Table (rows 1-9) ---
    # Row 1: table header label
    ws["A1"] = "Operating Expenses (Q1 2025)"
    ws["A1"].font = Font(bold=True, size=12)

    # Row 2: column headers
    ws["A2"] = "Expense Category"
    ws["B2"] = "Amount"
    ws["A2"].font = Font(bold=True)
    ws["B2"].font = Font(bold=True)

    # Rows 3-9: expense data (7 rows of realistic expense categories)
    expenses = [
        ("Salaries & Benefits",  142500),
        ("Rent & Utilities",      28400),
        ("Marketing & Advertising", 19750),
        ("Software Subscriptions", 8320),
        ("Travel & Conferences",   5600),
        ("Office Supplies",        2180),
        ("Miscellaneous",          3250),
    ]
    for r, (category, amount) in enumerate(expenses, 3):
        ws.cell(row=r, column=1, value=category)
        ws.cell(row=r, column=2, value=amount)

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18

    # --- Revenue Trend Table (rows 11-24) ---
    # Row 11: table header label
    ws["D11"] = "Monthly Revenue Trend (2024)"
    ws["D11"].font = Font(bold=True, size=12)

    # Row 12: column headers
    ws["D12"] = "Month"
    ws["E12"] = "Revenue"
    ws["D12"].font = Font(bold=True)
    ws["E12"].font = Font(bold=True)

    # Rows 13-24: 12 months of revenue data
    monthly_revenue = [
        ("January",   215400),
        ("February",  198700),
        ("March",     232600),
        ("April",     247800),
        ("May",       261300),
        ("June",      278900),
        ("July",      255100),
        ("August",    269400),
        ("September", 284700),
        ("October",   301200),
        ("November",  318500),
        ("December",  342800),
    ]
    for r, (month, revenue) in enumerate(monthly_revenue, 13):
        ws.cell(row=r, column=4, value=month)
        ws.cell(row=r, column=5, value=revenue)

    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
