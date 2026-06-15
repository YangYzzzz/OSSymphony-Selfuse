"""
Reward Script: Goal Seek scenario — determine deals needed to hit annual revenue target
Task ID: calc_sales_075
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): B3 contains formula =B1-B2 (Q4 Gap)
  Component 2 (0.2): B5 contains formula =ROUNDUP(B3/B4,0) (Deals Needed)
  Component 3 (0.5): B9:B15 contain sensitivity formulas =ROUNDUP($B$3/Ax,0)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_075'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: GoalSeek sheet must exist
    if 'GoalSeek' not in wb.sheetnames:
        print("FAIL: Sheet 'GoalSeek' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['GoalSeek']

    # Component 1: B3 contains formula =B1-B2 (Q4 Gap calculation) (0.3 points)
    try:
        b3_val = ws['B3'].value
        expected_b3 = '=B1-B2'
        if b3_val is not None and normalize_formula(b3_val) == normalize_formula(expected_b3):
            print(f"PASS: Component 1 — B3 has formula {b3_val} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Expected formula {expected_b3} in B3, found: {repr(b3_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B5 contains formula =ROUNDUP(B3/B4,0) (Deals Needed) (0.2 points)
    try:
        b5_val = ws['B5'].value
        expected_b5 = '=ROUNDUP(B3/B4,0)'
        if b5_val is not None and normalize_formula(b5_val) == normalize_formula(expected_b5):
            print(f"PASS: Component 2 — B5 has formula {b5_val} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Expected formula {expected_b5} in B5, found: {repr(b5_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B9:B15 contain sensitivity analysis formulas =ROUNDUP($B$3/Ax,0) (0.5 points)
    # Each correct formula earns 0.5/7 points (approximately 0.0714 each)
    try:
        sensitivity_score = 0.0
        expected_rows = {
            9: '=ROUNDUP($B$3/A9,0)',
            10: '=ROUNDUP($B$3/A10,0)',
            11: '=ROUNDUP($B$3/A11,0)',
            12: '=ROUNDUP($B$3/A12,0)',
            13: '=ROUNDUP($B$3/A13,0)',
            14: '=ROUNDUP($B$3/A14,0)',
            15: '=ROUNDUP($B$3/A15,0)',
        }
        correct_count = 0
        for row_num, expected_formula in expected_rows.items():
            cell_ref = f'B{row_num}'
            cell_val = ws[cell_ref].value
            if cell_val is not None and normalize_formula(cell_val) == normalize_formula(expected_formula):
                correct_count += 1
                print(f"  PASS: {cell_ref} has formula {cell_val}")
            else:
                print(f"  FAIL: {cell_ref} — Expected {expected_formula}, found: {repr(cell_val)}")

        if correct_count > 0:
            sensitivity_score = 0.5 * (correct_count / 7)
            print(f"PASS: Component 3 — {correct_count}/7 sensitivity formulas correct ({sensitivity_score:.4f} pts)")
            total_score += sensitivity_score
        else:
            print(f"FAIL: Component 3 — No sensitivity formulas found in B9:B15")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
