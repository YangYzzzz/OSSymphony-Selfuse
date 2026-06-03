"""
Initial Setup: Goal Seek scenario for Q4 deals needed
Task ID: calc_sales_075
Domain: libreoffice_calc

Creates a GoalSeek sheet with labels and raw data values.
Formulas in B3, B5, B9:B15 are left empty (task requires the agent to add them).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'GoalSeek'

    # --- Header styling ---
    header_font = Font(name='Calibri', size=11, bold=True)
    currency_fmt = '$#,##0'
    number_fmt = '#,##0'

    # --- Row 1: Annual Target ---
    ws['A1'] = 'Annual Target'
    ws['A1'].font = header_font
    ws['B1'] = 2000000
    ws['B1'].number_format = currency_fmt

    # --- Row 2: YTD Revenue ---
    ws['A2'] = 'YTD Revenue (Q1-Q3)'
    ws['A2'].font = header_font
    ws['B2'] = 1450000
    ws['B2'].number_format = currency_fmt

    # --- Row 3: Q4 Gap (empty - agent must add formula) ---
    ws['A3'] = 'Q4 Gap'
    ws['A3'].font = header_font
    # B3 intentionally left empty
    ws['B3'].number_format = currency_fmt

    # --- Row 4: Average Deal Size ---
    ws['A4'] = 'Average Deal Size'
    ws['A4'].font = header_font
    ws['B4'] = 55000
    ws['B4'].number_format = currency_fmt

    # --- Row 5: Deals Needed (empty - agent must add formula) ---
    ws['A5'] = 'Deals Needed in Q4'
    ws['A5'].font = header_font
    # B5 intentionally left empty
    ws['B5'].number_format = number_fmt

    # --- Row 7: Sensitivity Analysis header ---
    ws['A7'] = 'Sensitivity Analysis'
    ws['A7'].font = Font(name='Calibri', size=12, bold=True)

    # --- Row 8: Column headers for sensitivity table ---
    ws['A8'] = 'Avg Deal Size'
    ws['A8'].font = header_font
    ws['B8'] = 'Deals Needed'
    ws['B8'].font = header_font

    # --- Rows 9-15: Sensitivity data (avg deal sizes) ---
    deal_sizes = [40000, 45000, 50000, 55000, 60000, 65000, 70000]
    for i, size in enumerate(deal_sizes):
        row = 9 + i
        ws.cell(row=row, column=1, value=size).number_format = currency_fmt
        # B9:B15 intentionally left empty - agent must add formulas

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
