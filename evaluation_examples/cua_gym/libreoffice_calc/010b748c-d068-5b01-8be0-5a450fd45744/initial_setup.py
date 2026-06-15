"""
Initial Setup: Apply Color Scale conditional formatting to performance scores
Task ID: calc_gfl_026
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Q4 Review'

    # --- Headers ---
    headers = ['Employee ID', 'Name', 'Team', 'Performance Score', 'Rank', 'Review Date', 'Reviewer']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Employee Data (39 rows, rows 2-40) ---
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'David', 'Priya', 'James', 'Yuki', 'Carlos',
        'Amara', 'Thomas', 'Lin', 'Robert', 'Fatima', 'Kevin', 'Ingrid',
        'Alejandro', 'Mei', 'Patrick', 'Zara', 'Henrik', 'Aisha', 'Benjamin',
        'Sofia', 'Daniel', 'Nkechi', 'William', 'Rania', 'Oliver', 'Chiara',
        'Rajesh', 'Emma', 'Lucas', 'Hana', 'Victor', 'Leila', 'Nathan',
        'Ananya', 'Gregory', 'Simone'
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrova', 'Kim', 'Sharma', 'O\'Brien', 'Tanaka', 'Rivera',
        'Okafor', 'Mueller', 'Zhang', 'Williams', 'Al-Rashidi', 'Park', 'Svensson',
        'Gonzalez', 'Wu', 'Sullivan', 'Hassan', 'Andersen', 'Diallo', 'Foster',
        'Moretti', 'Nakamura', 'Eze', 'Barrett', 'Khoury', 'Thompson', 'Bianchi',
        'Patel', 'Larsson', 'Ferreira', 'Watanabe', 'Dubois', 'Mansour', 'Clarke',
        'Reddy', 'Novak', 'Delacroix'
    ]
    teams = ['Engineering', 'Marketing', 'Sales', 'Finance', 'Operations', 'HR', 'Product', 'Design']
    ranks = ['Junior', 'Mid-Level', 'Senior', 'Lead', 'Principal']
    reviewers = [
        'Diana Ross', 'Michael Torres', 'Christine Wang', 'Robert Hayes',
        'Sandra Mitchell', 'Andrew Chen', 'Laura Martinez', 'Steven Park'
    ]

    random.seed(42)

    # Generate performance scores ensuring we have a good spread from 1 to 10
    scores = []
    for i in range(39):
        # Ensure good distribution across the 1-10 range
        score = (i % 10) + 1
        scores.append(score)
    random.shuffle(scores)

    months = ['01', '02', '03', '10', '11', '12']

    for i in range(39):
        row = i + 2
        emp_id = f'EMP-{1001 + i}'
        name = f'{first_names[i]} {last_names[i]}'
        team = teams[i % len(teams)]
        score = scores[i]
        rank = ranks[i % len(ranks)]
        month = months[i % len(months)]
        day = random.randint(1, 28)
        review_date = f'2025-{month}-{day:02d}'
        reviewer = reviewers[i % len(reviewers)]

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=team)
        ws.cell(row=row, column=4, value=score)
        ws.cell(row=row, column=5, value=rank)
        ws.cell(row=row, column=6, value=review_date)
        ws.cell(row=row, column=7, value=reviewer)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
