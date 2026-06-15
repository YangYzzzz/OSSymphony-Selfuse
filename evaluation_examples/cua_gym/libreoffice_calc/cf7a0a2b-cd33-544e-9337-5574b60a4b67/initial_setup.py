"""
Initial Setup: Scholarship applicant data in Excel + blank PDF form template on desktop
Task ID: osworld_multi_apps_excel_pdf_form_007
Domain: libreoffice_calc (multi-app: calc + pdf)
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_excel_pdf_form_007'
DESKTOP = '/home/user/Desktop'

FONT_REGULAR = '/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf'
FONT_BOLD    = '/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf'
FONT_ITALIC  = '/usr/share/fonts/truetype/liberation2/LiberationSans-Italic.ttf'

# ── helpers ──────────────────────────────────────────────────────────────────

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# ── Step 1: Create scholarship_candidates.xlsx ────────────────────────────────

def create_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Scholarship Candidates"

    # Column headers
    headers = [
        "ApplicantName", "StudentID", "GPA",
        "FinancialNeed", "EssaySubmitted",
        "RecommendationLetters", "ScholarshipType"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name="Calibri", size=11)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4",
                                fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Realistic applicant data
    data = [
        ["Emily Hartwell",   "S2024-0041", 3.87, "High",   "Y", 3, "Merit-Based Excellence"],
        ["Daniel Okonkwo",   "S2024-0089", 3.45, "Medium", "Y", 2, "Community Leadership"],
        ["Priya Nambiar",    "S2024-0133", 3.92, "Low",    "Y", 3, "STEM Achievement"],
        ["Carlos Vega",      "S2024-0178", 3.20, "High",   "N", 1, "First-Generation Scholar"],
        ["Sophia Lindqvist", "S2024-0215", 3.68, "Medium", "Y", 2, "Global Studies Award"],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Adjust column widths
    col_widths = [20, 14, 6, 14, 16, 22, 26]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = w

    output = f'{WORKDIR}/scholarship_candidates.xlsx'
    wb.save(output)
    print(f'Excel file created: {output}')
    return output


# ── Step 2: Create scholarship_form_template.pdf ──────────────────────────────

def create_pdf_template():
    """
    Build a blank scholarship application PDF form.
    Fields: Applicant Name, Student ID, GPA, Financial Need (radio: High/Medium/Low),
            Essay Submitted (checkbox Y/N), Recommendation Letters (count), Scholarship Type.
    """
    from fpdf import FPDF

    class ScholarshipForm(FPDF):
        def _setup_fonts(self):
            self.add_font('LiberationSans', style='',  fname=FONT_REGULAR)
            self.add_font('LiberationSans', style='B', fname=FONT_BOLD)
            self.add_font('LiberationSans', style='I', fname=FONT_ITALIC)

        def header(self):
            self.set_font("LiberationSans", "B", 16)
            self.cell(0, 10, "SCHOLARSHIP APPLICATION FORM", align="C",
                      new_x="LMARGIN", new_y="NEXT")
            self.set_font("LiberationSans", "", 10)
            self.cell(0, 6, "Office of Financial Aid and Scholarships", align="C",
                      new_x="LMARGIN", new_y="NEXT")
            self.ln(4)
            self.set_line_width(0.5)
            self.line(10, self.get_y(), 200, self.get_y())
            self.ln(6)

        def footer(self):
            self.set_y(-15)
            self.set_font("LiberationSans", "I", 8)
            self.cell(0, 10,
                      "Page 1 of 1 | For office use only - do not alter this form",
                      align="C")

    pdf = ScholarshipForm()
    pdf.set_margins(15, 20, 15)
    pdf._setup_fonts()
    pdf.add_page()

    # ── Section 1: Applicant Information ──
    pdf.set_font("LiberationSans", "B", 12)
    pdf.cell(0, 8, "SECTION 1: Applicant Information",
             new_x="LMARGIN", new_y="NEXT")
    pdf.set_line_width(0.3)
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(4)

    def labeled_line(label, line_width=140):
        pdf.set_font("LiberationSans", "B", 10)
        pdf.cell(50, 7, label + ":", new_x="RIGHT", new_y="TOP")
        pdf.set_font("LiberationSans", "", 10)
        x_start = pdf.get_x()
        y_pos = pdf.get_y()
        pdf.cell(line_width, 7, "", new_x="LMARGIN", new_y="NEXT")
        pdf.set_line_width(0.2)
        pdf.line(x_start, y_pos + 7, x_start + line_width, y_pos + 7)
        pdf.ln(2)

    labeled_line("Applicant Name")
    labeled_line("Student ID")
    labeled_line("GPA (on 4.0 scale)")
    labeled_line("Scholarship Type")

    pdf.ln(4)

    # ── Section 2: Financial Need ──
    pdf.set_font("LiberationSans", "B", 12)
    pdf.cell(0, 8, "SECTION 2: Financial Need Level",
             new_x="LMARGIN", new_y="NEXT")
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(4)

    pdf.set_font("LiberationSans", "", 10)
    pdf.cell(0, 6, "Please indicate the applicant's financial need level:",
             new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    def radio_option(label):
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.set_line_width(0.4)
        pdf.circle(x + 3, y + 3, 3)
        pdf.set_xy(x + 10, y)
        pdf.set_font("LiberationSans", "", 11)
        pdf.cell(60, 7, label, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

    radio_option("High")
    radio_option("Medium")
    radio_option("Low")

    pdf.ln(4)

    # ── Section 3: Supporting Materials ──
    pdf.set_font("LiberationSans", "B", 12)
    pdf.cell(0, 8, "SECTION 3: Supporting Materials",
             new_x="LMARGIN", new_y="NEXT")
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(4)

    pdf.set_font("LiberationSans", "", 10)
    pdf.cell(0, 6, "Essay Submitted:", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)

    def checkbox_option(label):
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.set_line_width(0.4)
        pdf.rect(x + 2, y, 5, 5)
        pdf.set_xy(x + 12, y)
        pdf.set_font("LiberationSans", "", 11)
        pdf.cell(50, 6, label, new_x="LMARGIN", new_y="NEXT")
        pdf.ln(1)

    checkbox_option("Yes")
    checkbox_option("No")

    pdf.ln(4)

    pdf.set_font("LiberationSans", "B", 10)
    pdf.cell(60, 7, "Number of Recommendation Letters:", new_x="RIGHT", new_y="TOP")
    pdf.set_font("LiberationSans", "", 10)
    x_start = pdf.get_x()
    y_pos = pdf.get_y()
    pdf.cell(40, 7, "", new_x="LMARGIN", new_y="NEXT")
    pdf.set_line_width(0.2)
    pdf.line(x_start, y_pos + 7, x_start + 40, y_pos + 7)
    pdf.ln(8)

    # ── Section 4: Office Use Only ──
    pdf.set_font("LiberationSans", "B", 12)
    pdf.cell(0, 8, "SECTION 4: Office Use Only", new_x="LMARGIN", new_y="NEXT")
    pdf.line(15, pdf.get_y(), 195, pdf.get_y())
    pdf.ln(4)
    pdf.set_font("LiberationSans", "", 10)
    labeled_line("Reviewer Name")
    labeled_line("Decision")
    labeled_line("Date Reviewed")

    output = f'{DESKTOP}/scholarship_form_template.pdf'
    os.makedirs(DESKTOP, exist_ok=True)
    pdf.output(output)
    print(f'PDF template created: {output}')
    return output


# ── Step 3: Main ──────────────────────────────────────────────────────────────

def main():
    os.makedirs(WORKDIR, exist_ok=True)

    excel_path = create_excel()
    create_pdf_template()

    # GUI-ready startup: open Calc with the candidates file
    launch_gui(f'libreoffice --calc "{excel_path}"', delay_sec=2.0)
    print('GUI_READY: LibreOffice Calc opened with scholarship_candidates.xlsx')


main()
