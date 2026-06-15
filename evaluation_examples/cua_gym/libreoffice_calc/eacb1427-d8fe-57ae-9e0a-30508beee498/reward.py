"""
Reward Script: Sales forecast with best/most likely/worst case scenarios
Task ID: calc_sales_049
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Year 1 formulas in C4:C6 = $B$1*(1+growth_rate)
  Component 2 (0.35): Year 2 formulas in D4:D6 = C*(1+growth_rate)
  Component 3 (0.30): Year 3 formulas in E4:E6 = D*(1+growth_rate)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_049'


def normalize_formula(f):
    """Normalize a formula for comparison: uppercase, strip spaces, remove leading =."""
    if not isinstance(f, str):
        return ''
    f = f.strip().upper().replace(' ', '')
    if f.startswith('='):
        f = f[1:]
    return f


def is_year1_formula(formula_str, row):
    """
    Check if formula computes B1*(1+growth_rate) for the given row.
    Acceptable patterns:
      =$B$1*(1+B<row>)
      =B1*(1+B<row>)
      =$B$1*(1+$B$<row>)
      and variations with order/grouping
    """
    f = normalize_formula(formula_str)
    if not f:
        return False
    # Check that it references B1 (or $B$1) and B<row> (or $B$<row>)
    # and contains multiplication and (1+...)
    # Pattern: $B$1*(1+B<row>) or B$1*(1+$B$<row>) etc.
    # Flexible: look for B1 ref and B<row> ref and structure (1+...)
    b1_ref = bool(re.search(r'\$?B\$?1', f))
    growth_ref = bool(re.search(r'\$?B\$?' + str(row), f))
    has_multiply = '*' in f
    has_one_plus = bool(re.search(r'\(1\+', f) or re.search(r'\+1\)', f))
    return b1_ref and growth_ref and has_multiply and has_one_plus


def is_compound_growth_formula(formula_str, prev_col, row):
    """
    Check if formula computes <prev_col><row>*(1+B<row>).
    E.g., =C4*(1+B4), =D5*(1+B5), etc.
    """
    f = normalize_formula(formula_str)
    if not f:
        return False
    prev_ref = bool(re.search(r'\$?' + prev_col + r'\$?' + str(row), f))
    growth_ref = bool(re.search(r'\$?B\$?' + str(row), f))
    has_multiply = '*' in f
    has_one_plus = bool(re.search(r'\(1\+', f) or re.search(r'\+1\)', f))
    return prev_ref and growth_ref and has_multiply and has_one_plus


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Forecast' sheet exists
    if 'Forecast' not in wb.sheetnames:
        print("FAIL: 'Forecast' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Forecast']

    # Component 1: Year 1 formulas in C4:C6 (0.35 points)
    # C4 = $B$1*(1+B4), C5 = $B$1*(1+B5), C6 = $B$1*(1+B6)
    try:
        year1_pass = 0
        for row in [4, 5, 6]:
            cell_val = ws.cell(row=row, column=3).value  # column C
            if is_year1_formula(cell_val, row):
                print(f"PASS: C{row} has correct Year 1 formula: {cell_val}")
                year1_pass += 1
            else:
                print(f"FAIL: C{row} expected Year 1 formula ($B$1*(1+B{row})), found: {repr(cell_val)}")
        if year1_pass == 3:
            print(f"PASS: Component 1 -- All Year 1 formulas correct (0.35 pts)")
            total_score += 0.35
        elif year1_pass > 0:
            partial = round(0.35 * year1_pass / 3, 4)
            print(f"PARTIAL: Component 1 -- {year1_pass}/3 Year 1 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No Year 1 formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Year 2 formulas in D4:D6 (0.35 points)
    # D4 = C4*(1+B4), D5 = C5*(1+B5), D6 = C6*(1+B6)
    try:
        year2_pass = 0
        for row in [4, 5, 6]:
            cell_val = ws.cell(row=row, column=4).value  # column D
            if is_compound_growth_formula(cell_val, 'C', row):
                print(f"PASS: D{row} has correct Year 2 formula: {cell_val}")
                year2_pass += 1
            else:
                print(f"FAIL: D{row} expected Year 2 formula (C{row}*(1+B{row})), found: {repr(cell_val)}")
        if year2_pass == 3:
            print(f"PASS: Component 2 -- All Year 2 formulas correct (0.35 pts)")
            total_score += 0.35
        elif year2_pass > 0:
            partial = round(0.35 * year2_pass / 3, 4)
            print(f"PARTIAL: Component 2 -- {year2_pass}/3 Year 2 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No Year 2 formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Year 3 formulas in E4:E6 (0.30 points)
    # E4 = D4*(1+B4), E5 = D5*(1+B5), E6 = D6*(1+B6)
    try:
        year3_pass = 0
        for row in [4, 5, 6]:
            cell_val = ws.cell(row=row, column=5).value  # column E
            if is_compound_growth_formula(cell_val, 'D', row):
                print(f"PASS: E{row} has correct Year 3 formula: {cell_val}")
                year3_pass += 1
            else:
                print(f"FAIL: E{row} expected Year 3 formula (D{row}*(1+B{row})), found: {repr(cell_val)}")
        if year3_pass == 3:
            print(f"PASS: Component 3 -- All Year 3 formulas correct (0.30 pts)")
            total_score += 0.30
        elif year3_pass > 0:
            partial = round(0.30 * year3_pass / 3, 4)
            print(f"PARTIAL: Component 3 -- {year3_pass}/3 Year 3 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No Year 3 formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
