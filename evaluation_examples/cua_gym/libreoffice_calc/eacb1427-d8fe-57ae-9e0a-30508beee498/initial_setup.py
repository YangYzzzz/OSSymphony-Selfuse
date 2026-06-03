"""
Initial Setup: Sales forecast with scenario analysis
Task ID: calc_sales_049
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Forecast ---
    ws = wb.active
    ws.title = 'Forecast'

    # Header styling
    header_font = Font(name='Calibri', size=12, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    currency_fmt = '#,##0'
    pct_fmt = '0%'

    # A1: Current Annual Revenue label, B1: value
    ws['A1'] = 'Current Annual Revenue'
    ws['A1'].font = Font(name='Calibri', size=12, bold=True)
    ws['B1'] = 1200000
    ws['B1'].number_format = '$#,##0'
    ws['B1'].font = Font(name='Calibri', size=12, bold=True)

    # Row 2: blank separator

    # Row 3: Table headers
    headers = {
        'A3': 'Scenario',
        'B3': 'Growth Rate',
        'C3': 'Year 1',
        'D3': 'Year 2',
        'E3': 'Year 3',
    }
    for coord, label in headers.items():
        cell = ws[coord]
        cell.value = label
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Row 4: Worst Case
    ws['A4'] = 'Worst Case'
    ws['A4'].font = Font(name='Calibri', size=11)
    ws['B4'] = 0.05
    ws['B4'].number_format = pct_fmt

    # Row 5: Most Likely
    ws['A5'] = 'Most Likely'
    ws['A5'].font = Font(name='Calibri', size=11)
    ws['B5'] = 0.15
    ws['B5'].number_format = pct_fmt

    # Row 6: Best Case
    ws['A6'] = 'Best Case'
    ws['A6'].font = Font(name='Calibri', size=11)
    ws['B6'] = 0.25
    ws['B6'].number_format = pct_fmt

    # C4:E6 intentionally left empty - that's the task for the agent

    # Column widths for readability
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    # Add a second sheet with supporting reference data for realism
    ws2 = wb.create_sheet('Historical Data')
    hist_headers = ['Year', 'Revenue', 'Growth %', 'Region']
    for c, h in enumerate(hist_headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    historical = [
        [2020, 850000, 0.08, 'North America'],
        [2021, 920000, 0.082, 'North America'],
        [2022, 1005000, 0.092, 'North America'],
        [2023, 1080000, 0.075, 'North America'],
        [2024, 1150000, 0.065, 'North America'],
        [2025, 1200000, 0.043, 'North America'],
    ]
    for r, row_data in enumerate(historical, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            if c == 2:
                cell.number_format = '$#,##0'
            elif c == 3:
                cell.number_format = '0.0%'

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 15
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
