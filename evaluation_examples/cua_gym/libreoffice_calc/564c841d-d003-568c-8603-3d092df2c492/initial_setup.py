"""
Initial Setup: KPI Dashboard layout template
Task ID: calc_gg2_007
Domain: libreoffice_calc

Creates a workbook with:
- A blank 'Dashboard' sheet (active)
- A 'Data' sheet with 6 KPI metrics and monthly trend data
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Data (KPI source data) ---
    ws_data = wb.active
    ws_data.title = 'Data'

    # KPI metrics and values (A1:B7)
    ws_data['A1'] = 'Metric'
    ws_data['B1'] = 'Value'
    ws_data['A1'].font = Font(bold=True)
    ws_data['B1'].font = Font(bold=True)

    kpi_data = [
        ['Revenue', 1285400],
        ['Gross Margin', 0.423],
        ['Customer Acquisition Cost', 187.50],
        ['Monthly Active Users', 34200],
        ['Churn Rate', 0.032],
        ['Net Promoter Score', 72],
    ]
    for r, (metric, value) in enumerate(kpi_data, 2):
        ws_data.cell(row=r, column=1, value=metric)
        ws_data.cell(row=r, column=2, value=value)

    # Monthly trend data (D1:H7)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May']
    ws_data['D1'] = months[0]
    ws_data['E1'] = months[1]
    ws_data['F1'] = months[2]
    ws_data['G1'] = months[3]
    ws_data['H1'] = months[4]
    for col_idx in range(4, 9):
        ws_data.cell(row=1, column=col_idx).font = Font(bold=True)

    trend_data = [
        [1050000, 1120000, 1180000, 1230000, 1285400],   # Revenue
        [0.398, 0.405, 0.412, 0.418, 0.423],             # Gross Margin
        [210.00, 205.00, 198.50, 192.00, 187.50],        # CAC
        [28500, 30100, 31800, 33000, 34200],              # MAU
        [0.045, 0.042, 0.038, 0.035, 0.032],             # Churn Rate
        [65, 67, 69, 71, 72],                             # NPS
    ]
    for r, row_vals in enumerate(trend_data, 2):
        for c, val in enumerate(row_vals, 4):
            ws_data.cell(row=r, column=c, value=val)

    # Set column widths on Data sheet
    ws_data.column_dimensions['A'].width = 28
    ws_data.column_dimensions['B'].width = 16
    for col_letter in ['D', 'E', 'F', 'G', 'H']:
        ws_data.column_dimensions[col_letter].width = 14

    # --- Sheet 2: Dashboard (blank, active) ---
    ws_dash = wb.create_sheet('Dashboard')
    # Leave it completely blank - the agent must build the dashboard
    wb.active = wb.sheetnames.index('Dashboard')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
