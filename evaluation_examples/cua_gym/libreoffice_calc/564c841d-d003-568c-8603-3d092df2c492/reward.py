"""
Reward Script: KPI Dashboard Layout in LibreOffice Calc
Task ID: calc_gg2_007
Domain: libreoffice_calc
Scoring:
  Component 1: Merged title row A1:J1 with bold title text (0.20)
  Component 2: Bordered summary table in A3:D8 with data (0.25)
  Component 3: Line chart embedded in the Dashboard sheet (0.25)
  Component 4: DRAFT image/shape in upper-right area (0.20)
  Component 5: Print area set on Dashboard sheet (0.10)
"""

import os
import zipfile
import xml.etree.ElementTree as ET

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_007'


def persist_app_state(domain: str):
    """Try to save any unsaved LibreOffice changes."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify KPI dashboard task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        from openpyxl.cell.cell import MergedCell
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dashboard sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dashboard']

    # Component 1: Merged title row A1:J1 with bold title text (0.20 points)
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'J1' in r for r in merged_ranges)
        # Also accept wider/narrower merges that include A1 through at least several columns
        if not has_merge:
            # Check if A1 is part of any merge range spanning multiple columns
            for mr in ws.merged_cells.ranges:
                if mr.min_row == 1 and mr.max_row == 1 and mr.min_col == 1 and mr.max_col >= 5:
                    has_merge = True
                    break

        title_val = ws['A1'].value
        has_title = title_val is not None and len(str(title_val).strip()) > 0
        is_bold = ws['A1'].font.bold == True
        font_size = ws['A1'].font.size

        if has_merge and has_title and is_bold:
            print(f"PASS: Component 1 — Merged title '{title_val}', bold={is_bold}, size={font_size} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — merge={has_merge}, title='{title_val}', bold={is_bold}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Bordered summary table in A3:D8 with data (0.25 points)
    try:
        # Check that A3:D8 has data (at least header row and some data rows)
        has_data = False
        data_cell_count = 0
        for row in ws.iter_rows(min_row=3, max_row=8, min_col=1, max_col=4):
            for cell in row:
                if not isinstance(cell, MergedCell) and cell.value is not None:
                    data_cell_count += 1

        # Need at least header row (4 cells) + some data = ~8 cells minimum
        has_data = data_cell_count >= 8

        # Check borders — outer cells should have thick borders, inner should have thin
        has_borders = False
        # Check a corner cell (A3) for thick left and thick top
        a3_border = ws['A3'].border
        d8_border = ws['D8'].border
        if (a3_border.left.style in ('thick', 'medium') and
            a3_border.top.style in ('thick', 'medium') and
            d8_border.right.style in ('thick', 'medium') and
            d8_border.bottom.style in ('thick', 'medium')):
            has_borders = True

        # Also check an inner border (B4) for thin
        b4_border = ws['B4'].border
        has_inner_borders = (b4_border.left.style == 'thin' and b4_border.top.style == 'thin')

        if has_data and has_borders and has_inner_borders:
            print(f"PASS: Component 2 — Table A3:D8 with {data_cell_count} data cells, thick outer + thin inner borders (0.25 pts)")
            total_score += 0.25
        elif has_data and (has_borders or has_inner_borders):
            print(f"PARTIAL: Component 2 — Data present but border structure incomplete (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — data_cells={data_cell_count}, outer_borders={has_borders}, inner_borders={has_inner_borders}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Line chart embedded in Dashboard sheet (0.25 points)
    try:
        charts = ws._charts
        has_chart = len(charts) >= 1

        if has_chart:
            chart = charts[0]
            chart_class = type(chart).__name__
            is_line = 'Line' in chart_class
            has_series = len(chart.series) > 0

            # Check chart title
            has_title = False
            if chart.title is not None:
                has_title = True

            if is_line and has_series and has_title:
                print(f"PASS: Component 3 — LineChart with {len(chart.series)} series and title (0.25 pts)")
                total_score += 0.25
            elif has_series:
                # Partial: chart exists but may not be line type or missing title
                print(f"PARTIAL: Component 3 — Chart present ({chart_class}) with {len(chart.series)} series, line={is_line}, title={has_title} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — Chart class={chart_class}, series={len(chart.series)}, title={has_title}")
        else:
            print(f"FAIL: Component 3 — No charts found on Dashboard sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: DRAFT image/shape in upper-right area (0.20 points)
    # The golden file uses an embedded image for the DRAFT rectangle.
    # We check for either: (a) a drawing image/shape in the xlsx, or (b) openpyxl images
    try:
        has_draft_element = False

        # Method 1: Check openpyxl images
        if hasattr(ws, '_images') and len(ws._images) > 0:
            has_draft_element = True
            print(f"  Found {len(ws._images)} image(s) via openpyxl")

        # Method 2: Parse the drawing XML for shapes or pictures beyond the chart
        if not has_draft_element:
            with zipfile.ZipFile(file_path, 'r') as z:
                drawing_files = [n for n in z.namelist() if 'drawing' in n.lower() and n.endswith('.xml')]
                for df in drawing_files:
                    content = z.read(df).decode()
                    ns_xdr = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
                    ns_a = 'http://schemas.openxmlformats.org/drawingml/2006/main'
                    root = ET.fromstring(content)

                    # Count anchors - chart is one, any additional anchor with pic or sp is the DRAFT element
                    anchors = root.findall(f'{{{ns_xdr}}}oneCellAnchor') + root.findall(f'{{{ns_xdr}}}twoCellAnchor')
                    pic_or_sp_count = 0
                    for anchor in anchors:
                        pics = anchor.findall(f'{{{ns_xdr}}}pic')
                        shapes = anchor.findall(f'{{{ns_xdr}}}sp')
                        if pics or shapes:
                            pic_or_sp_count += 1
                            # Check if shape has DRAFT text
                            for sp in shapes:
                                texts = sp.findall(f'.//{{{ns_a}}}t')
                                text_content = ''.join([t.text or '' for t in texts])
                                if 'DRAFT' in text_content.upper():
                                    print(f"  Found shape with text: '{text_content}'")

                    if pic_or_sp_count >= 1:
                        has_draft_element = True
                        print(f"  Found {pic_or_sp_count} pic/shape element(s) in drawing XML")

        # Method 3: Check for embedded media files (images)
        if not has_draft_element:
            with zipfile.ZipFile(file_path, 'r') as z:
                media_files = [n for n in z.namelist() if 'media/' in n.lower()]
                if len(media_files) > 0:
                    has_draft_element = True
                    print(f"  Found {len(media_files)} media file(s)")

        if has_draft_element:
            print(f"PASS: Component 4 — DRAFT annotation element found (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — No DRAFT image/shape found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Print area set on Dashboard sheet (0.10 points)
    try:
        print_area = ws.print_area
        has_print_area = print_area is not None and len(str(print_area)) > 0

        if has_print_area:
            print(f"PASS: Component 5 — Print area set: {print_area} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No print area defined on Dashboard")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
