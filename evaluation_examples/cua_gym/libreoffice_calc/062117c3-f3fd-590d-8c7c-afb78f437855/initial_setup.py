"""
Initial Setup: Hide columns D, E, F on 'Confidential' sheet
Task ID: calc_gg1_041
Domain: libreoffice_calc

Creates a staff directory spreadsheet with a 'Confidential' sheet containing
employee data across 7 columns. Columns D (Salary), E (SSN), F (Home Address)
are all VISIBLE in the initial state. A second 'Overview' sheet provides context.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Confidential ---
    ws = wb.active
    ws.title = 'Confidential'

    headers = ['Name', 'Department', 'Title', 'Salary', 'SSN', 'Home Address', 'Work Phone']
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Realistic employee data
    data = [
        ['Sarah Chen', 'Engineering', 'Senior Software Engineer', 128500, '331-44-8821', '1247 Maple Dr, Austin TX 78704', '(512) 555-0134'],
        ['Marcus Johnson', 'Marketing', 'Marketing Manager', 97200, '442-55-9932', '890 Oak Lane, Austin TX 78731', '(512) 555-0267'],
        ['Emily Rodriguez', 'Finance', 'Financial Analyst', 82000, '553-66-1143', '2301 Pine St Apt 4B, Austin TX 78745', '(512) 555-0398'],
        ['David Kim', 'Engineering', 'DevOps Lead', 115000, '664-77-2254', '456 Cedar Blvd, Round Rock TX 78664', '(512) 555-0421'],
        ['Rachel Patel', 'Human Resources', 'HR Director', 135000, '775-88-3365', '7823 Elm Way, Austin TX 78703', '(512) 555-0553'],
        ['James O\'Brien', 'Sales', 'Regional Sales Director', 110000, '886-99-4476', '1100 Birch Ave, Cedar Park TX 78613', '(512) 555-0687'],
        ['Lisa Nakamura', 'Engineering', 'QA Engineer', 89500, '997-00-5587', '3402 Willow Ct, Austin TX 78748', '(512) 555-0719'],
        ['Carlos Mendez', 'Operations', 'Operations Coordinator', 67500, '108-11-6698', '5614 Spruce Ln, Pflugerville TX 78660', '(512) 555-0842'],
        ['Amanda Foster', 'Finance', 'Accounting Manager', 98000, '219-22-7709', '928 Ash St, Georgetown TX 78628', '(512) 555-0975'],
        ['Michael Zhang', 'Engineering', 'Principal Architect', 155000, '320-33-8810', '4501 Juniper Rd, Austin TX 78735', '(512) 555-1108'],
        ['Priya Sharma', 'Marketing', 'Content Strategist', 76000, '431-44-9921', '2789 Poplar Dr, Austin TX 78723', '(512) 555-1231'],
        ['Thomas Williams', 'Sales', 'Account Executive', 85000, '542-55-0032', '6100 Magnolia Blvd, Leander TX 78641', '(512) 555-1364'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 4:  # Salary column - number format
                cell.number_format = '$#,##0'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 36
    ws.column_dimensions['G'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Overview ---
    ws2 = wb.create_sheet('Overview')
    ws2['A1'] = 'Staff Directory Overview'
    ws2['A1'].font = Font(bold=True, size=14)
    ws2['A3'] = 'Department'
    ws2['B3'] = 'Headcount'
    ws2['A3'].font = Font(bold=True)
    ws2['B3'].font = Font(bold=True)

    dept_counts = [
        ['Engineering', 4],
        ['Marketing', 2],
        ['Finance', 2],
        ['Human Resources', 1],
        ['Sales', 2],
        ['Operations', 1],
    ]
    for r, row_data in enumerate(dept_counts, 4):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2['A11'] = 'Last Updated: 2025-03-28'
    ws2['A12'] = 'Maintained by: HR Department'
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
