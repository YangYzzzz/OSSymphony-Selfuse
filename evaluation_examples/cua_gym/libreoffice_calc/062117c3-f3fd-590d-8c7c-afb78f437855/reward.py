"""
Reward Script: Hide columns D, E, and F on the 'Confidential' sheet
Task ID: calc_gg1_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Column D is hidden on Confidential sheet
  Component 2 (0.30): Column E is hidden on Confidential sheet
  Component 3 (0.30): Column F is hidden on Confidential sheet
  Component 4 (0.10): Data in D, E, F is preserved (not deleted)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_041'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that columns D, E, F are hidden on the 'Confidential' sheet
    while data is preserved. Progressive scoring 0.0 - 1.0.
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Confidential' sheet must exist
    if 'Confidential' not in wb.sheetnames:
        print("CRITICAL: 'Confidential' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Confidential']

    # Component 1: Column D is hidden (0.30 points)
    try:
        col_d = ws.column_dimensions.get('D')
        if col_d and col_d.hidden:
            print(f"PASS: Component 1 -- Column D is hidden (0.30 pts)")
            total_score += 0.30
        else:
            hidden_val = col_d.hidden if col_d else 'no dimension'
            print(f"FAIL: Component 1 -- Column D hidden={hidden_val}, expected True")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Column E is hidden (0.30 points)
    try:
        col_e = ws.column_dimensions.get('E')
        if col_e and col_e.hidden:
            print(f"PASS: Component 2 -- Column E is hidden (0.30 pts)")
            total_score += 0.30
        else:
            hidden_val = col_e.hidden if col_e else 'no dimension'
            print(f"FAIL: Component 2 -- Column E hidden={hidden_val}, expected True")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Column F is hidden (0.30 points)
    try:
        col_f = ws.column_dimensions.get('F')
        if col_f and col_f.hidden:
            print(f"PASS: Component 3 -- Column F is hidden (0.30 pts)")
            total_score += 0.30
        else:
            hidden_val = col_f.hidden if col_f else 'no dimension'
            print(f"FAIL: Component 3 -- Column F hidden={hidden_val}, expected True")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Columns are hidden AND data is preserved (0.10 points)
    # This compound check only awards points when the hiding was done correctly
    # AND the underlying data was not deleted. It anchors to the task change
    # (hidden columns) so it fails on initial_env where columns are not hidden.
    try:
        col_d_hidden = ws.column_dimensions.get('D') and ws.column_dimensions['D'].hidden
        col_e_hidden = ws.column_dimensions.get('E') and ws.column_dimensions['E'].hidden
        col_f_hidden = ws.column_dimensions.get('F') and ws.column_dimensions['F'].hidden
        any_hidden = col_d_hidden or col_e_hidden or col_f_hidden

        d1 = ws['D1'].value  # Should be 'Salary'
        e1 = ws['E1'].value  # Should be 'SSN'
        f1 = ws['F1'].value  # Should be 'Home Address'
        d2 = ws['D2'].value  # Should be a salary number
        e2 = ws['E2'].value  # Should be an SSN string
        f2 = ws['F2'].value  # Should be an address string

        all_present = all(v is not None for v in [d1, e1, f1, d2, e2, f2])
        headers_ok = (
            str(d1).strip() == 'Salary'
            and str(e1).strip() == 'SSN'
            and str(f1).strip() == 'Home Address'
        )

        if any_hidden and all_present and headers_ok:
            print(f"PASS: Component 4 -- Columns hidden AND data preserved "
                  f"(D1={d1}, E1={e1}, F1={f1}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 -- any_hidden={any_hidden}, "
                  f"all_present={all_present}, headers_ok={headers_ok}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist unsaved GUI state, then verify
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
