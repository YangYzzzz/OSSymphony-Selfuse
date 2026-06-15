"""
Initial Setup: Shipment consolidation analysis - create data and empty summary sheet
Task ID: calc_ops_093
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 'Shipments' ---
    ws1 = wb.active
    ws1.title = 'Shipments'

    headers = ['ID', 'Destination', 'Weight', 'Cost']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    data = [
        ['S01', 'Dallas',  500,  750],
        ['S02', 'Phoenix', 300,  510],
        ['S03', 'Dallas',  750, 1050],
        ['S05', 'Dallas',  200,  320],
        ['S06', 'Phoenix', 600,  960],
        ['S07', 'Atlanta', 400,  680],
        ['S08', 'Atlanta', 350,  595],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 10

    # --- Sheet 'Summary' ---
    ws2 = wb.create_sheet('Summary')

    summary_headers = ['City', '# Shipments', 'Total Weight', 'Total Cost', 'Avg Cost/kg']
    for col, h in enumerate(summary_headers, 1):
        ws2.cell(row=1, column=col, value=h)

    # City names only - B:E columns left empty for the agent to fill
    ws2.cell(row=2, column=1, value='Dallas')
    ws2.cell(row=3, column=1, value='Phoenix')
    ws2.cell(row=4, column=1, value='Atlanta')

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 14
    ws2.column_dimensions['D'].width = 12
    ws2.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
