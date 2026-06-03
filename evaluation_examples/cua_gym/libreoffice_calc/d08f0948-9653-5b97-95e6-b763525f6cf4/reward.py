"""
Reward Script: Shipment consolidation analysis with pivot-style summary
Task ID: calc_ops_093
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): COUNTIF formulas in B2:B4 for shipment counts
  Component 2 (0.30): SUMIF formulas in C2:C4 and D2:D4 for total weight/cost
  Component 3 (0.20): Division formulas in E2:E4 for avg cost/kg
  Component 4 (0.20): Computed values match ground truth (via data_only)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_093'

# Ground truth from task context
GROUND_TRUTH = {
    'Dallas':  {'row': 2, 'count': 3, 'weight': 1450, 'cost': 2120, 'avg': 1.462},
    'Phoenix': {'row': 3, 'count': 2, 'weight': 900,  'cost': 1470, 'avg': 1.633},
    'Atlanta': {'row': 4, 'count': 2, 'weight': 750,  'cost': 1275, 'avg': 1.700},
}


def is_countif_formula(val, city_ref):
    """Check if value is a COUNTIF formula referencing the Shipments destination column."""
    if not isinstance(val, str):
        return False
    v = val.upper().replace(" ", "")
    # Accept patterns like =COUNTIF(Shipments.B:B,A2) or =COUNTIF(Shipments!B:B,A2)
    # or =COUNTIF(Shipments.B$1:B$8,A2) etc.
    return 'COUNTIF(' in v and ('SHIPMENTS' in v or 'B:B' in v or 'B$' in v)


def is_sumif_formula(val, expected_col):
    """Check if value is a SUMIF formula referencing expected column (C or D)."""
    if not isinstance(val, str):
        return False
    v = val.upper().replace(" ", "")
    # Must contain SUMIF and reference the expected column from Shipments
    if 'SUMIF(' not in v:
        return False
    # Check it references the right sum range column
    col_letter = expected_col.upper()
    # Accept Shipments.C:C or Shipments!C:C or just C:C patterns
    return (f'{col_letter}:{col_letter}' in v or f'{col_letter}$' in v or
            f'SHIPMENTS.{col_letter}' in v or f'SHIPMENTS!{col_letter}' in v)


def is_division_formula(val, row):
    """Check if value is a division formula like =D2/C2."""
    if not isinstance(val, str):
        return False
    v = val.upper().replace(" ", "")
    # Accept =D{row}/C{row} or variations
    return '/' in v and f'D{row}' in v and f'C{row}' in v


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: COUNTIF formulas for shipment counts in B2:B4 (0.30 points)
    # 0.10 per city row
    try:
        comp1_score = 0.0
        for city, info in GROUND_TRUTH.items():
            r = info['row']
            val = ws.cell(row=r, column=2).value  # Column B
            city_ref = f'A{r}'
            if is_countif_formula(val, city_ref):
                print(f"PASS: Component 1 — {city} (B{r}) has COUNTIF formula: {val} (0.10 pts)")
                comp1_score += 0.10
            else:
                print(f"FAIL: Component 1 — {city} (B{r}) expected COUNTIF formula, found: {val}")
        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SUMIF formulas for total weight (C) and total cost (D) in rows 2-4 (0.30 points)
    # 0.05 per cell (6 cells total = 0.30)
    try:
        comp2_score = 0.0
        for city, info in GROUND_TRUTH.items():
            r = info['row']
            # Check C column (Total Weight) — SUMIF on Shipments.C column
            val_c = ws.cell(row=r, column=3).value
            if is_sumif_formula(val_c, 'C'):
                print(f"PASS: Component 2 — {city} (C{r}) has SUMIF formula for weight: {val_c} (0.05 pts)")
                comp2_score += 0.05
            else:
                print(f"FAIL: Component 2 — {city} (C{r}) expected SUMIF formula for weight, found: {val_c}")

            # Check D column (Total Cost) — SUMIF on Shipments.D column
            val_d = ws.cell(row=r, column=4).value
            if is_sumif_formula(val_d, 'D'):
                print(f"PASS: Component 2 — {city} (D{r}) has SUMIF formula for cost: {val_d} (0.05 pts)")
                comp2_score += 0.05
            else:
                print(f"FAIL: Component 2 — {city} (D{r}) expected SUMIF formula for cost, found: {val_d}")
        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Division formulas in E2:E4 for avg cost/kg (0.20 points)
    # 0.067 per city (~0.20 total)
    try:
        comp3_score = 0.0
        for city, info in GROUND_TRUTH.items():
            r = info['row']
            val_e = ws.cell(row=r, column=5).value
            if is_division_formula(val_e, r):
                print(f"PASS: Component 3 — {city} (E{r}) has division formula: {val_e} (0.067 pts)")
                comp3_score += 0.067
            else:
                print(f"FAIL: Component 3 — {city} (E{r}) expected division formula (D{r}/C{r}), found: {val_e}")
        # Round to avoid floating point issues
        comp3_score = round(min(comp3_score, 0.20), 3)
        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Verify computed values match ground truth (0.20 points)
    # Use data_only=True to get cached values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['Summary']
        comp4_score = 0.0
        checks_passed = 0
        total_checks = 9  # 3 cities x 3 values (count, weight, cost)

        for city, info in GROUND_TRUTH.items():
            r = info['row']

            # Check count (B column)
            b_val = ws_data.cell(row=r, column=2).value
            if b_val is not None:
                try:
                    if abs(float(b_val) - info['count']) < 0.01:
                        checks_passed += 1
                        print(f"PASS: Component 4 — {city} count = {b_val} (expected {info['count']})")
                    else:
                        print(f"FAIL: Component 4 — {city} count = {b_val} (expected {info['count']})")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4 — {city} count not numeric: {b_val}")
            else:
                print(f"INFO: Component 4 — {city} count (B{r}) cached value is None (file may not have been opened in Calc)")

            # Check weight (C column)
            c_val = ws_data.cell(row=r, column=3).value
            if c_val is not None:
                try:
                    if abs(float(c_val) - info['weight']) < 0.1:
                        checks_passed += 1
                        print(f"PASS: Component 4 — {city} weight = {c_val} (expected {info['weight']})")
                    else:
                        print(f"FAIL: Component 4 — {city} weight = {c_val} (expected {info['weight']})")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4 — {city} weight not numeric: {c_val}")
            else:
                print(f"INFO: Component 4 — {city} weight (C{r}) cached value is None")

            # Check cost (D column)
            d_val = ws_data.cell(row=r, column=4).value
            if d_val is not None:
                try:
                    if abs(float(d_val) - info['cost']) < 0.1:
                        checks_passed += 1
                        print(f"PASS: Component 4 — {city} cost = {d_val} (expected {info['cost']})")
                    else:
                        print(f"FAIL: Component 4 — {city} cost = {d_val} (expected {info['cost']})")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4 — {city} cost not numeric: {d_val}")
            else:
                print(f"INFO: Component 4 — {city} cost (D{r}) cached value is None")

        # Award proportional score for cached values that match
        # If all cached values are None (file never opened in Calc), award full 0.20 if formulas are correct
        non_none_count = sum(
            1 for city, info in GROUND_TRUTH.items()
            for col in [2, 3, 4]
            if ws_data.cell(row=info['row'], column=col).value is not None
        )

        if non_none_count == 0:
            # No cached values available — formulas were already verified in Components 1-3
            # Award full component 4 score if all formula components scored > 0
            if total_score >= 0.75:
                comp4_score = 0.20
                print("INFO: Component 4 — No cached values (file not opened in Calc), but all formulas correct. Awarding 0.20 pts.")
            else:
                print("INFO: Component 4 — No cached values and some formulas missing. Awarding 0.0 pts.")
        else:
            comp4_score = round((checks_passed / total_checks) * 0.20, 3)
            print(f"INFO: Component 4 — {checks_passed}/{total_checks} value checks passed ({comp4_score} pts)")

        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
