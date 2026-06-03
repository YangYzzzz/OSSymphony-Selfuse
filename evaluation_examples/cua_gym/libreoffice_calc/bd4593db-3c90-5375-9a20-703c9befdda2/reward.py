"""
Reward Script: Set up project issues log with severity/category dropdowns,
               Days Open formula, Week Raised formula, COUNTIFS summary table,
               and grouped bar chart showing issues raised vs resolved per week.
Task ID: calc_ops_project_issues_log_065
Domain: libreoffice_calc
Scoring:
  - Component 1: Data validation dropdowns on Category (C), Severity (D), Status (G)  — 0.30 points
  - Component 2: Days Open IF formula in I2:I71                                         — 0.25 points
  - Component 3: Week Raised TEXT formula in J2:J71                                    — 0.20 points
  - Component 4: IssueSummary COUNTIFS formulas (B2:C5)                                — 0.15 points
  - Component 5: IssueSummary Total formula (D2:D5 = B+C)                              — 0.10 points
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_project_issues_log_065'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition: Sheets exist ---
    if 'IssuesLog' not in wb.sheetnames:
        print("CRITICAL: Sheet 'IssuesLog' not found")
        print("REWARD: 0.0")
        return 0.0
    if 'IssueSummary' not in wb.sheetnames:
        print("CRITICAL: Sheet 'IssueSummary' not found")
        print("REWARD: 0.0")
        return 0.0

    ws_log = wb['IssuesLog']
    ws_sum = wb['IssueSummary']

    # -------------------------------------------------------------------------
    # Component 1: Data Validation Dropdowns (0.30 points total)
    # Three dropdowns required:
    #   - C2:C71: Category — Technical, Process, Resource, Scope, External
    #   - D2:D71: Severity — Critical, High, Medium, Low
    #   - G2:G71: Status   — Open, In Progress, Resolved, Closed, Cancelled
    # Each dropdown is worth 0.10 points.
    # -------------------------------------------------------------------------
    try:
        dvs = ws_log.data_validations.dataValidation

        # Build a dictionary mapping column ranges to their validations
        dv_map = {}
        for dv in dvs:
            if dv.type == 'list' and dv.formula1:
                sqref_str = str(dv.sqref)
                dv_map[sqref_str] = dv.formula1

        # Helper to find a data validation covering a specific range and containing expected values
        def find_dv_for_range(target_range, expected_values):
            """
            Check if any data validation covers the target range and contains the expected values.
            Expected values is a list of strings.
            """
            for sqref_str, formula1 in dv_map.items():
                # Check if the target range appears in the sqref
                if target_range in sqref_str or sqref_str in target_range:
                    # Remove surrounding quotes from formula1 if present
                    f1_clean = formula1.strip('"')
                    values_in_dv = [v.strip() for v in f1_clean.split(',')]
                    # Check all expected values are present
                    if all(ev in values_in_dv for ev in expected_values):
                        return True, formula1
            return False, None

        # Category dropdown: C2:C71
        cat_found, cat_formula = find_dv_for_range('C2:C71',
            ['Technical', 'Process', 'Resource', 'Scope', 'External'])
        if cat_found:
            print(f"PASS: Component 1a — Category dropdown C2:C71 found (formula: {cat_formula}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1a — Category dropdown C2:C71 not found. DVs present: {list(dv_map.items())}")

        # Severity dropdown: D2:D71
        sev_found, sev_formula = find_dv_for_range('D2:D71',
            ['Critical', 'High', 'Medium', 'Low'])
        if sev_found:
            print(f"PASS: Component 1b — Severity dropdown D2:D71 found (formula: {sev_formula}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1b — Severity dropdown D2:D71 not found. DVs present: {list(dv_map.items())}")

        # Status dropdown: G2:G71
        stat_found, stat_formula = find_dv_for_range('G2:G71',
            ['Open', 'In Progress', 'Resolved', 'Closed', 'Cancelled'])
        if stat_found:
            print(f"PASS: Component 1c — Status dropdown G2:G71 found (formula: {stat_formula}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1c — Status dropdown G2:G71 not found. DVs present: {list(dv_map.items())}")

    except Exception as e:
        print(f"ERROR: Component 1 (Data Validations) — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Days Open IF Formula in I2:I71 (0.25 points)
    # Each cell in I2:I71 should contain: =IF(G<row>="Resolved",H<row>-B<row>,TODAY()-B<row>)
    # Check that:
    #   - At least 90% of cells (63 of 70) have an IF formula referencing G, H, B columns
    # Award 0.25 if all rows have the formula, 0.15 if >= 50% have it.
    # -------------------------------------------------------------------------
    try:
        if_count = 0
        total_rows = 70  # rows 2-71

        for row in range(2, 72):
            cell_val = ws_log.cell(row=row, column=9).value  # Column I
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Check for IF formula with Resolved check and date subtraction
                if (val_upper.startswith('=IF(') and
                    'RESOLVED' in val_upper and
                    f'B{row}' in cell_val):
                    if_count += 1

        ratio = if_count / total_rows
        if ratio >= 0.95:
            print(f"PASS: Component 2 — Days Open IF formula in {if_count}/{total_rows} rows (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.50:
            partial = 0.15
            print(f"PARTIAL: Component 2 — Days Open IF formula in {if_count}/{total_rows} rows (0.15 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Days Open IF formula only in {if_count}/{total_rows} rows")

    except Exception as e:
        print(f"ERROR: Component 2 (Days Open formula) — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Week Raised TEXT Formula in J2:J71 (0.20 points)
    # Each cell in J2:J71 should contain a formula that calculates the week start date.
    # Examples: =WEEKNUM(B2) or =TEXT(B2-WEEKDAY(B2,2)+1,"DD-MMM")
    # Check that at least 90% of cells have a week-related formula referencing column B.
    # -------------------------------------------------------------------------
    try:
        week_count = 0
        total_rows = 70

        for row in range(2, 72):
            cell_val = ws_log.cell(row=row, column=10).value  # Column J
            if cell_val and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Accept WEEKNUM(B<row>) or TEXT(...WEEKDAY...B<row>...) patterns
                has_weeknum = (f'WEEKNUM(B{row})' in val_upper or
                               f'WEEKNUM(B{row},' in val_upper)
                has_text_week = ('TEXT(' in val_upper and
                                 'WEEKDAY' in val_upper and
                                 f'B{row}' in cell_val)
                has_isoweeknum = (f'ISOWEEKNUM(B{row})' in val_upper)
                if has_weeknum or has_text_week or has_isoweeknum:
                    week_count += 1

        ratio = week_count / total_rows
        if ratio >= 0.95:
            print(f"PASS: Component 3 — Week Raised formula in {week_count}/{total_rows} rows (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.50:
            partial = 0.12
            print(f"PARTIAL: Component 3 — Week Raised formula in {week_count}/{total_rows} rows (0.12 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Week Raised formula only in {week_count}/{total_rows} rows")
            # Show sample to help debug
            sample = ws_log.cell(row=2, column=10).value
            print(f"  Sample J2 value: {repr(sample)}")

    except Exception as e:
        print(f"ERROR: Component 3 (Week Raised formula) — {e}")

    # -------------------------------------------------------------------------
    # Component 4: IssueSummary COUNTIFS Formulas (0.15 points)
    # B2:B5 should contain COUNTIFS for Open issues by severity
    # C2:C5 should contain COUNTIFS for Resolved issues by severity
    # Check that all 8 cells (B2:B5 and C2:C5) have COUNTIFS referencing IssuesLog
    # -------------------------------------------------------------------------
    try:
        countifs_count = 0
        total_cells = 8  # B2:B5 and C2:C5

        for row in range(2, 6):
            # B column (Open counts)
            b_val = ws_sum.cell(row=row, column=2).value
            if b_val and isinstance(b_val, str):
                val_upper = b_val.upper().replace(' ', '')
                if 'COUNTIFS(' in val_upper and 'ISSUESLOG' in val_upper:
                    countifs_count += 1

            # C column (Resolved counts)
            c_val = ws_sum.cell(row=row, column=3).value
            if c_val and isinstance(c_val, str):
                val_upper = c_val.upper().replace(' ', '')
                if 'COUNTIFS(' in val_upper and 'ISSUESLOG' in val_upper:
                    countifs_count += 1

        if countifs_count == total_cells:
            print(f"PASS: Component 4 — COUNTIFS formulas in all {total_cells} cells (B2:C5) (0.15 pts)")
            total_score += 0.15
        elif countifs_count >= 4:
            partial = 0.08
            print(f"PARTIAL: Component 4 — COUNTIFS formulas in {countifs_count}/{total_cells} cells (0.08 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — COUNTIFS formulas in only {countifs_count}/{total_cells} cells")

    except Exception as e:
        print(f"ERROR: Component 4 (IssueSummary COUNTIFS) — {e}")

    # -------------------------------------------------------------------------
    # Component 5: IssueSummary Total Formula in D2:D5 (0.10 points)
    # D2:D5 should each contain a formula that adds Open Count + Resolved Count
    # e.g., =B2+C2
    # -------------------------------------------------------------------------
    try:
        total_formula_count = 0
        for row in range(2, 6):
            d_val = ws_sum.cell(row=row, column=4).value
            if d_val and isinstance(d_val, str):
                val_stripped = d_val.replace(' ', '')
                # Accept =B<row>+C<row> or =SUM(B<row>:C<row>) patterns
                is_sum_formula = (
                    val_stripped == f'=B{row}+C{row}' or
                    val_stripped.upper() == f'=SUM(B{row}:C{row})' or
                    # Allow in either order
                    val_stripped == f'=C{row}+B{row}'
                )
                if is_sum_formula:
                    total_formula_count += 1

        if total_formula_count == 4:
            print(f"PASS: Component 5 — Total formula in all 4 cells (D2:D5) (0.10 pts)")
            total_score += 0.10
        elif total_formula_count >= 2:
            partial = 0.05
            print(f"PARTIAL: Component 5 — Total formula in {total_formula_count}/4 cells (0.05 pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Total formula in only {total_formula_count}/4 cells")
            for row in range(2, 6):
                d_val = ws_sum.cell(row=row, column=4).value
                print(f"  D{row}: {repr(d_val)}")

    except Exception as e:
        print(f"ERROR: Component 5 (IssueSummary Total formula) — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
