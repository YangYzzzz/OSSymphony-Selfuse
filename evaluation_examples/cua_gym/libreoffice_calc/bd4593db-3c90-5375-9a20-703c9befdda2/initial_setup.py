"""
Initial Setup: Project Issues Log with severity and category data (no dropdowns/formulas yet)
Task ID: calc_ops_project_issues_log_065
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_issues_log_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: IssuesLog
    # ------------------------------------------------------------------ #
    ws = wb.active
    ws.title = 'IssuesLog'

    # Headers row 1
    headers = [
        'Issue ID',     # A
        'Date Raised',  # B
        'Category',     # C  -- needs dropdown (NOT yet added)
        'Severity',     # D  -- needs dropdown (NOT yet added)
        'Description',  # E
        'Assigned To',  # F
        'Status',       # G  -- needs dropdown (NOT yet added)
        'Date Resolved',# H
        'Days Open',    # I  -- empty (formula to be added)
        'Week Raised',  # J  -- empty (formula to be added)
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 70 issues: A, B, C, D, E, F, G filled; H filled for resolved items
    # I and J are intentionally left EMPTY (task asks to calculate them)
    # C, D, G are filled with plain text values (NOT as dropdown validated cells)
    import datetime

    categories = ['Technical', 'Process', 'Resource', 'Scope', 'External']
    severities = ['Critical', 'High', 'Medium', 'Low']
    statuses   = ['Open', 'In Progress', 'Resolved', 'Closed', 'Cancelled']

    assignees = [
        'Alice Nguyen', 'Ben Carter', 'Clara Mills', 'David Park',
        'Eva Hernandez', 'Frank Liu', 'Grace Kim', 'Hiro Tanaka',
        'Isabel Reyes', 'James Okafor'
    ]

    descriptions = [
        'Login page throws 500 error on invalid credentials',
        'Report generation timeout for large datasets',
        'CI pipeline fails on feature branch merges',
        'Onboarding checklist missing step for new contractors',
        'Database migration script does not handle null values',
        'Email notifications sent in wrong timezone',
        'User permissions reset after password change',
        'API rate limiting not enforced for internal services',
        'PDF export crops footer content on A4 paper',
        'Search index not updated after bulk imports',
        'Dashboard metrics lag by 15 minutes during peak hours',
        'Audit log missing entries for bulk deletions',
        'Two-factor authentication bypass via deep link',
        'Mobile app crashes on iOS 17.3 during file upload',
        'Scheduled task fails silently after server restart',
        'CSV import rejects valid date formats from EU locales',
        'Tooltip text overflows container on small screens',
        'Webhook retry logic sends duplicate events',
        'Config file ignored when environment variable is set',
        'Session timeout not enforced on shared workstations',
        'Pagination breaks when filter is applied then cleared',
        'Sorting by date column ignores time component',
        'Attachment preview fails for HEIC image files',
        'Role assignment does not propagate to sub-groups',
        'Drag-and-drop reorder not saved on page refresh',
        'Memory leak in background sync service',
        'Error message exposes internal stack trace to users',
        'Backup job skips files larger than 2 GB',
        'Localization strings missing for Spanish locale',
        'Dark mode toggle reverts after browser refresh',
        'Print layout breaks for tables wider than 12 columns',
        'Offline mode does not sync changes on reconnect',
        'Analytics script blocked by default browser settings',
        'User avatar upload fails for PNG files over 1 MB',
        'Keyboard shortcut conflicts with OS-level binding',
        'Comment thread collapses unexpectedly on scroll',
        'QR code scanner returns wrong encoding for UTF-8 data',
        'Custom domain SSL certificate not auto-renewed',
        'Import wizard skips last row of data silently',
        'Gantt chart dates shift by one day in UTC-5 timezone',
        'Template variables not escaped in email subject line',
        'Guest users can access private project boards via URL',
        'Notification badge count resets on page navigation',
        'Calendar integration fails for recurring events with exceptions',
        'Bulk export ZIP file exceeds 4 GB limit on Windows',
        'Auto-save conflicts when same document opened in two tabs',
        'Filter dropdown list does not load when cache is cold',
        'Heatmap colours inverted for negative values',
        'Admin panel accessible without MFA on VPN network',
        'Workflow approval step skipped when approver is also requester',
        'Chart legend truncates long series names',
        'Inline edit mode loses focus on validation error',
        'Dependency graph cycles cause infinite loop in analysis',
        'Row highlighting deselects on table sort action',
        'API token expiry not communicated to integration partners',
        'Custom report columns reorder unexpectedly after save',
        'Webhook secret not validated on retry requests',
        'Tag autocomplete suggests archived tags',
        'Copy-paste of multi-line cells loses line breaks in CSV',
        'Colour-blind mode missing for status indicators',
        'Push notification deep link navigates to wrong screen',
        'Realtime collaboration cursor disappears after 30 seconds',
        'SAML SSO breaks when assertion attribute order changes',
        'Currency symbol missing in PDF invoice export',
        'Timeline zoom out button disabled at minimum zoom level',
        'Tooltip shows raw Markdown syntax instead of rendered text',
        'Sidebar collapses when window width is exactly 1024 px',
        'Graph edge labels overlap at high node density',
        'Undo history cleared when switching between project views',
        'Search results do not highlight keywords in attachment text',
    ]

    base_date = datetime.date(2025, 1, 6)  # a Monday

    rows_data = []
    for i in range(70):
        issue_num = i + 1
        issue_id = f'PROJ-{issue_num:03d}'

        # Date raised: stagger over ~20 weeks
        weeks_offset = i // 4
        day_offset   = (i % 4) * 2
        date_raised  = base_date + datetime.timedelta(weeks=weeks_offset, days=day_offset)

        cat      = categories[i % len(categories)]
        severity = severities[i % len(severities)]
        desc     = descriptions[i % len(descriptions)]
        assignee = assignees[i % len(assignees)]

        # Alternate statuses: roughly 40% Open, 20% In Progress, 30% Resolved, 5% Closed, 5% Cancelled
        status_cycle = ['Open', 'In Progress', 'Resolved', 'Open', 'Resolved',
                        'Open', 'In Progress', 'Resolved', 'Closed', 'Cancelled']
        status = status_cycle[i % len(status_cycle)]

        # Date resolved: only for Resolved/Closed
        date_resolved = None
        if status in ('Resolved', 'Closed'):
            date_resolved = date_raised + datetime.timedelta(days=3 + (i % 12))

        # I (Days Open) and J (Week Raised) left EMPTY — task asks agent to add formula
        rows_data.append([
            issue_id, date_raised, cat, severity, desc, assignee,
            status, date_resolved, None, None
        ])

    for r_idx, row in enumerate(rows_data, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    # Format date columns
    for r in range(2, 72):
        ws.cell(row=r, column=2).number_format = 'YYYY-MM-DD'
        cell_h = ws.cell(row=r, column=8)
        if cell_h.value is not None:
            cell_h.number_format = 'YYYY-MM-DD'

    # Column widths
    col_widths = [12, 14, 14, 12, 50, 18, 14, 16, 12, 14]
    import string
    alpha = list(string.ascii_uppercase)
    for i, w in enumerate(col_widths):
        ws.column_dimensions[alpha[i]].width = w

    ws.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Sheet 2: IssueSummary
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('IssueSummary')

    summary_headers = ['Severity', 'Open Count', 'Resolved Count', 'Total']
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    severities_list = ['Critical', 'High', 'Medium', 'Low']
    for r, sev in enumerate(severities_list, 2):
        ws2.cell(row=r, column=1, value=sev)
        # B, C, D intentionally EMPTY — task asks agent to add COUNTIFS formulas

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
