"""
Reward Script: Track IEP goal progress for special education students
Task ID: calc_edu_special_ed_iep_tracker_061
Domain: libreoffice_calc
Scoring:
  Component 1: Column G (Goal Avg) — AVERAGE formulas for rows 2-46 (0.35 pts)
  Component 2: Column H (Regression Flag) — IF(OR(...)) formulas for rows 2-46 (0.30 pts)
  Component 3: Summary rows B49:B63 — AVERAGEIF formulas per student (0.20 pts)
  Component 4: Conditional formatting — orange fill for 'Regression' in H2:H46 (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_special_ed_iep_tracker_061'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: IEPTracker sheet must exist
    if 'IEPTracker' not in wb.sheetnames:
        print(f"CRITICAL: 'IEPTracker' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['IEPTracker']

    # Component 1: Column G Goal Avg — AVERAGE(Cn:Fn) formulas in rows 2-46 (0.35 points)
    # This FAILS on initial (G is None) → should PASS on golden (has AVERAGE formulas)
    try:
        avg_formula_count = 0
        avg_format_count = 0
        expected_avg_rows = 45  # rows 2 through 46

        for row in range(2, 47):
            cell_g = ws.cell(row=row, column=7)
            val = cell_g.value
            if val is not None and isinstance(val, str):
                # Check it's an AVERAGE formula referencing the same row's C:F columns
                val_upper = val.upper().replace(" ", "")
                expected_formula = f"=AVERAGE(C{row}:F{row})"
                if val_upper == expected_formula.upper().replace(" ", ""):
                    avg_formula_count += 1
                    # Also check number format is 1 decimal place
                    if cell_g.number_format in ('0.0', '#,##0.0', '0.0;-0.0'):
                        avg_format_count += 1

        if avg_formula_count == expected_avg_rows:
            print(f"PASS: Component 1 — All {expected_avg_rows} AVERAGE formulas present in column G (0.35 pts)")
            total_score += 0.35
        elif avg_formula_count >= expected_avg_rows * 0.8:
            # Partial credit: at least 80% of formulas present
            partial = 0.20
            print(f"PARTIAL: Component 1 — {avg_formula_count}/{expected_avg_rows} AVERAGE formulas in column G ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {avg_formula_count}/{expected_avg_rows} AVERAGE formulas found in column G")
    except Exception as e:
        print(f"ERROR: Component 1 (Goal Avg formulas) — {e}")

    # Component 2: Column H Regression Flag — IF(OR(...)) formulas detecting score drops (0.30 points)
    # This FAILS on initial (H is None) → should PASS on golden (has IF/OR formulas)
    try:
        reg_formula_count = 0
        expected_reg_rows = 45  # rows 2 through 46

        for row in range(2, 47):
            cell_h = ws.cell(row=row, column=8)
            val = cell_h.value
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(" ", "")
                # Check it contains IF, OR, and "REGRESSION" text
                if ('=IF(' in val_upper and 'OR(' in val_upper and
                        'REGRESSION' in val_upper and 'PROGRESS' in val_upper):
                    # Verify it checks consecutive quarter drops (D<C, E<D, F<E pattern)
                    col_refs = [
                        f'D{row}<C{row}',
                        f'E{row}<D{row}',
                        f'F{row}<E{row}'
                    ]
                    checks_pass = all(ref in val_upper for ref in col_refs)
                    if checks_pass:
                        reg_formula_count += 1

        if reg_formula_count == expected_reg_rows:
            print(f"PASS: Component 2 — All {expected_reg_rows} Regression Flag formulas present in column H (0.30 pts)")
            total_score += 0.30
        elif reg_formula_count >= expected_reg_rows * 0.8:
            partial = 0.18
            print(f"PARTIAL: Component 2 — {reg_formula_count}/{expected_reg_rows} Regression Flag formulas in column H ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {reg_formula_count}/{expected_reg_rows} Regression Flag formulas found in column H")
    except Exception as e:
        print(f"ERROR: Component 2 (Regression Flag formulas) — {e}")

    # Component 3: Student summary rows B49:B63 — AVERAGEIF formulas (0.20 points)
    # This FAILS on initial (B49:B63 are None) → should PASS on golden (has AVERAGEIF formulas)
    try:
        averageif_count = 0
        expected_students = 15  # rows 49 through 63

        for row in range(49, 64):
            cell_b = ws.cell(row=row, column=2)
            val = cell_b.value
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(" ", "")
                # Check for AVERAGEIF referencing column A range and column G range
                if ('=AVERAGEIF(' in val_upper and
                        '$A$2:$A$46' in val_upper and
                        '$G$2:$G$46' in val_upper):
                    averageif_count += 1

        if averageif_count == expected_students:
            print(f"PASS: Component 3 — All {expected_students} AVERAGEIF formulas present in B49:B63 (0.20 pts)")
            total_score += 0.20
        elif averageif_count >= expected_students * 0.8:
            partial = 0.12
            print(f"PARTIAL: Component 3 — {averageif_count}/{expected_students} AVERAGEIF formulas in B49:B63 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {averageif_count}/{expected_students} AVERAGEIF formulas in B49:B63")
    except Exception as e:
        print(f"ERROR: Component 3 (AVERAGEIF formulas) — {e}")

    # Component 4: Conditional formatting — orange fill for 'Regression' in column H (0.15 points)
    # This FAILS on initial (no CF rules) → should PASS on golden (has orange fill CF on H2:H46)
    try:
        cf_found = False
        orange_fill_found = False
        regression_formula_found = False

        cf_rules = ws.conditional_formatting
        for cf in cf_rules:
            cf_range_str = str(cf)
            # Check if CF applies to column H range
            if 'H' in cf_range_str:
                for rule in cf.rules:
                    # Check for formula-based rule matching 'Regression'
                    if rule.type == 'expression' and rule.formula:
                        formula_upper = rule.formula[0].upper().replace(' ', '')
                        if 'REGRESSION' in formula_upper:
                            regression_formula_found = True
                    # Check for orange fill
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            # Orange: FFFFA500 or similar orange hex
                            if fill_color and fill_color.upper() in ('FFFFA500', 'FFA500'):
                                orange_fill_found = True
                        except Exception:
                            pass
                    if regression_formula_found and orange_fill_found:
                        cf_found = True
                        break
            if cf_found:
                break

        if cf_found:
            print(f"PASS: Component 4 — Conditional formatting with orange fill for 'Regression' in column H (0.15 pts)")
            total_score += 0.15
        elif regression_formula_found:
            # CF rule exists but color may differ
            partial = 0.08
            print(f"PARTIAL: Component 4 — Regression CF rule found but orange fill color not confirmed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No orange conditional formatting for 'Regression' found in column H")
    except Exception as e:
        print(f"ERROR: Component 4 (Conditional formatting) — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
