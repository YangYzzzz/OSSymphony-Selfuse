"""
Initial Setup: IEP Goal Progress Tracker for Special Education Students
Task ID: calc_edu_special_ed_iep_tracker_061
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_special_ed_iep_tracker_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'IEPTracker'

    # --- Headers (Row 1) ---
    headers = ['Student', 'Goal', 'Q1 Score', 'Q2 Score', 'Q3 Score', 'Q4 Score', 'Goal Avg', 'Regression Flag']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- 15 students x 3 goals = 45 rows (rows 2-46) ---
    # Realistic special education student names and IEP goals
    students = [
        'Aiden Torres',
        'Brianna Mitchell',
        'Carlos Rivera',
        'Diana Patel',
        'Ethan Nguyen',
        'Fatima Hassan',
        'Gabriel Kim',
        'Hannah Brooks',
        'Isaiah Cooper',
        'Jasmine Ortega',
        'Kevin Pham',
        'Lily Sanderson',
        'Marcus Williams',
        'Nadia Flores',
        'Owen Tran',
    ]

    goals_by_student = [
        # Aiden Torres
        [
            ('Aiden Torres', 'Reading Comprehension - Identify main idea', 2, 2, 3, 3),
            ('Aiden Torres', 'Written Expression - Write 5-sentence paragraphs', 1, 2, 2, 3),
            ('Aiden Torres', 'Math - Solve two-step word problems', 2, 3, 3, 4),
        ],
        # Brianna Mitchell
        [
            ('Brianna Mitchell', 'Social Skills - Initiate peer interactions', 3, 3, 4, 4),
            ('Brianna Mitchell', 'Reading Fluency - Read 80 wpm with accuracy', 2, 3, 3, 3),
            ('Brianna Mitchell', 'Math - Add/subtract fractions', 1, 2, 2, 3),
        ],
        # Carlos Rivera
        [
            ('Carlos Rivera', 'Speech - Use complete sentences in conversation', 2, 2, 3, 4),
            ('Carlos Rivera', 'Reading - Decode multisyllabic words', 1, 2, 3, 3),
            ('Carlos Rivera', 'Behavior - Follow classroom rules independently', 3, 3, 3, 4),
        ],
        # Diana Patel
        [
            ('Diana Patel', 'Math - Multiply multi-digit numbers', 2, 3, 4, 4),
            ('Diana Patel', 'Writing - Use punctuation correctly', 2, 2, 3, 3),
            ('Diana Patel', 'Organization - Complete assignments on time', 1, 2, 2, 3),
        ],
        # Ethan Nguyen
        [
            ('Ethan Nguyen', 'Reading Comprehension - Draw inferences from text', 3, 3, 3, 4),
            ('Ethan Nguyen', 'Math - Solve equations with one variable', 2, 2, 3, 3),
            ('Ethan Nguyen', 'Social-Emotional - Manage frustration appropriately', 2, 3, 3, 4),
        ],
        # Fatima Hassan
        [
            ('Fatima Hassan', 'Reading - Identify story elements', 1, 2, 3, 3),
            ('Fatima Hassan', 'Math - Tell time to nearest minute', 2, 3, 3, 4),
            ('Fatima Hassan', 'Life Skills - Follow two-step directions', 3, 3, 4, 4),
        ],
        # Gabriel Kim
        [
            ('Gabriel Kim', 'Writing - Organize ideas in paragraph form', 2, 2, 2, 3),
            ('Gabriel Kim', 'Math - Count money and make change', 3, 4, 4, 4),
            ('Gabriel Kim', 'Behavior - Stay on task for 20 minutes', 2, 3, 3, 3),
        ],
        # Hannah Brooks
        [
            ('Hannah Brooks', 'Speech - Articulate /r/ sound in sentences', 2, 3, 3, 4),
            ('Hannah Brooks', 'Reading - Comprehend grade-level fiction', 1, 2, 2, 3),
            ('Hannah Brooks', 'Math - Identify geometric shapes and properties', 3, 3, 4, 4),
        ],
        # Isaiah Cooper
        [
            ('Isaiah Cooper', 'Reading Fluency - Read with prosody and expression', 2, 2, 3, 3),
            ('Isaiah Cooper', 'Math - Understand place value to thousands', 3, 3, 3, 4),
            ('Isaiah Cooper', 'Writing - Expand vocabulary in written work', 1, 2, 3, 3),
        ],
        # Jasmine Ortega
        [
            ('Jasmine Ortega', 'Social Skills - Cooperate in group activities', 3, 4, 4, 4),
            ('Jasmine Ortega', 'Reading - Use context clues to determine meaning', 2, 2, 3, 3),
            ('Jasmine Ortega', 'Math - Add/subtract decimals', 1, 2, 2, 3),
        ],
        # Kevin Pham
        [
            ('Kevin Pham', 'Behavior - Raise hand before speaking', 2, 3, 3, 4),
            ('Kevin Pham', 'Reading - Identify cause and effect relationships', 3, 3, 4, 4),
            ('Kevin Pham', 'Math - Solve problems involving measurement', 2, 2, 2, 3),
        ],
        # Lily Sanderson
        [
            ('Lily Sanderson', 'Speech - Use appropriate volume in classroom', 3, 3, 4, 4),
            ('Lily Sanderson', 'Reading Comprehension - Summarize informational text', 2, 2, 3, 3),
            ('Lily Sanderson', 'Math - Compare and order fractions', 1, 2, 3, 3),
        ],
        # Marcus Williams
        [
            ('Marcus Williams', 'Writing - Use transition words in paragraphs', 2, 3, 3, 4),
            ('Marcus Williams', 'Math - Solve division problems with remainders', 3, 3, 4, 4),
            ('Marcus Williams', 'Social-Emotional - Identify and express emotions', 2, 2, 3, 3),
        ],
        # Nadia Flores
        [
            ('Nadia Flores', 'Reading - Distinguish fact from opinion', 1, 2, 2, 3),
            ('Nadia Flores', 'Math - Identify patterns and complete sequences', 2, 3, 3, 4),
            ('Nadia Flores', 'Life Skills - Follow daily schedule independently', 3, 3, 3, 4),
        ],
        # Owen Tran
        [
            ('Owen Tran', 'Reading - Phonemic awareness and decoding', 2, 2, 3, 3),
            ('Owen Tran', 'Math - Add/subtract two-digit numbers with regrouping', 1, 2, 3, 4),
            ('Owen Tran', 'Behavior - Complete transitions without prompting', 3, 3, 4, 4),
        ],
    ]

    row = 2
    for student_goals in goals_by_student:
        for goal_row in student_goals:
            ws.cell(row=row, column=1, value=goal_row[0])  # Student
            ws.cell(row=row, column=2, value=goal_row[1])  # Goal
            ws.cell(row=row, column=3, value=goal_row[2])  # Q1 Score
            ws.cell(row=row, column=4, value=goal_row[3])  # Q2 Score
            ws.cell(row=row, column=5, value=goal_row[4])  # Q3 Score
            ws.cell(row=row, column=6, value=goal_row[5])  # Q4 Score
            # G (Goal Avg) and H (Regression Flag) left EMPTY intentionally
            row += 1

    # Row 47: blank separator
    # Row 48: Summary header
    ws.cell(row=47, column=1, value='')
    summary_header_student = ws.cell(row=48, column=1, value='Student')
    summary_header_progress = ws.cell(row=48, column=2, value='Overall Progress')
    summary_header_student.font = Font(bold=True)
    summary_header_progress.font = Font(bold=True)

    # Rows 49-63: student summaries (column B left empty — to be filled by AVERAGEIF)
    for i, student in enumerate(students):
        ws.cell(row=49 + i, column=1, value=student)
        # Column B left empty intentionally (to be calculated)

    # Column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
