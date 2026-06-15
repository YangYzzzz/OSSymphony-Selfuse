"""
Reward Script: Configure print selection to print only 'Summary' and 'Charts' sheets
Task ID: calc_mcp_078
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Charts sheet has tabSelected=True
  Component 2 (0.5): Exactly 2 sheets selected, and they are Summary and Charts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_078'


def persist_app_state(domain: str):
    """Best-effort save of any unsaved GUI state."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    The task requires selecting the 'Summary' and 'Charts' sheets for printing
    (i.e., they should have tabSelected=True in the sheet views), while 'Details'
    and 'Raw Data' should NOT be selected.

    We verify this by reading the tabSelected attribute from each sheet's view.

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: expected sheets must exist
    expected_sheets = ['Summary', 'Details', 'Charts', 'Raw Data']
    for sname in expected_sheets:
        if sname not in wb.sheetnames:
            print(f"CRITICAL: Expected sheet '{sname}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Helper: get tabSelected status for a sheet
    def is_tab_selected(sheet_name):
        """Return True if the sheet's first sheetView has tabSelected=True."""
        ws = wb[sheet_name]
        try:
            views = ws.views
            if views and views.sheetView:
                return bool(views.sheetView[0].tabSelected)
        except Exception as e:
            print(f"WARNING: Could not read tabSelected for '{sheet_name}': {e}")
        return False

    # Read all tab selection states
    tab_states = {}
    for sname in expected_sheets:
        tab_states[sname] = is_tab_selected(sname)
        print(f"INFO: Sheet '{sname}' tabSelected = {tab_states[sname]}")

    # Component 1: Charts sheet has tabSelected=True (0.5 points)
    # This is the key change from initial state where Charts was NOT selected.
    try:
        if tab_states['Charts']:
            print(f"PASS: Component 1 -- Charts sheet is tabSelected (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 -- Charts sheet is NOT tabSelected")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Exactly 2 sheets selected, and they are Summary and Charts (0.5 points)
    # This ensures the correct exclusive selection: only Summary + Charts, not Details or Raw Data.
    try:
        selected_sheets = [s for s in expected_sheets if tab_states[s]]
        selected_count = len(selected_sheets)
        correct_selection = (
            selected_count == 2
            and tab_states['Summary']
            and tab_states['Charts']
            and not tab_states['Details']
            and not tab_states['Raw Data']
        )
        if correct_selection:
            print(f"PASS: Component 2 -- Exactly 2 sheets selected: {selected_sheets} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 -- Expected exactly Summary+Charts selected, "
                  f"found {selected_count} selected: {selected_sheets}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
