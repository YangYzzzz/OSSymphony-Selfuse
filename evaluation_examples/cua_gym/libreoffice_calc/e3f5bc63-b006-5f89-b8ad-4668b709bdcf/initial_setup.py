"""
Initial Setup: Create workbook with Summary, Details, Charts, Raw Data sheets
Task ID: calc_mcp_078
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Common styles ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header(ws, row, num_cols):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_headers = ["Metric", "Q1 2025", "Q2 2025", "Q3 2025", "Q4 2025", "Annual Total"]
    for c, h in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=c, value=h)
    style_header(ws_summary, 1, len(summary_headers))

    summary_data = [
        ["Total Revenue", 245800, 312500, 287600, 356200, 1202100],
        ["Operating Costs", 128400, 145200, 139800, 152600, 566000],
        ["Gross Profit", 117400, 167300, 147800, 203600, 636100],
        ["Marketing Spend", 32000, 45000, 38000, 51000, 166000],
        ["R&D Investment", 55000, 58000, 62000, 65000, 240000],
        ["Net Income", 30400, 64300, 47800, 87600, 230100],
        ["Headcount", 42, 45, 48, 52, None],
        ["Customer Acquisition Cost", 185.50, 172.30, 168.90, 155.20, None],
        ["Customer Lifetime Value", 2450, 2680, 2710, 2890, None],
        ["Monthly Recurring Revenue", 81933, 104167, 95867, 118733, None],
        ["Churn Rate (%)", 3.2, 2.8, 2.5, 2.1, None],
        ["NPS Score", 72, 75, 78, 81, None],
    ]
    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_summary.cell(row=r, column=c, value=val)
            if c >= 2 and val is not None and isinstance(val, (int, float)):
                if "%" not in str(summary_data[r - 2][0]) and "NPS" not in str(summary_data[r - 2][0]) and "Headcount" not in str(summary_data[r - 2][0]):
                    cell.number_format = '$#,##0.00'

    ws_summary.column_dimensions["A"].width = 28
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws_summary.column_dimensions[col_letter].width = 15

    # === Sheet 2: Details ===
    ws_details = wb.create_sheet("Details")

    details_headers = ["Date", "Transaction ID", "Client", "Product", "Quantity", "Unit Price", "Total", "Region", "Sales Rep"]
    for c, h in enumerate(details_headers, 1):
        ws_details.cell(row=1, column=c, value=h)
    style_header(ws_details, 1, len(details_headers))

    details_data = [
        ["2025-01-08", "TXN-10241", "Meridian Corp", "Enterprise Suite", 3, 4500, 13500, "Northeast", "Sarah Chen"],
        ["2025-01-15", "TXN-10242", "Atlas Industries", "Pro License", 12, 850, 10200, "West Coast", "Marcus Johnson"],
        ["2025-02-03", "TXN-10243", "Pinnacle Group", "Enterprise Suite", 1, 4500, 4500, "Southeast", "Priya Sharma"],
        ["2025-02-14", "TXN-10244", "Vertex Solutions", "Starter Pack", 25, 200, 5000, "Midwest", "David Kim"],
        ["2025-03-01", "TXN-10245", "Nova Technologies", "Pro License", 8, 850, 6800, "West Coast", "Sarah Chen"],
        ["2025-03-22", "TXN-10246", "Summit Partners", "Enterprise Suite", 2, 4500, 9000, "Northeast", "Marcus Johnson"],
        ["2025-04-05", "TXN-10247", "Horizon Labs", "Starter Pack", 50, 200, 10000, "Southeast", "Priya Sharma"],
        ["2025-04-18", "TXN-10248", "Eclipse Dynamics", "Pro License", 5, 850, 4250, "Midwest", "David Kim"],
        ["2025-05-10", "TXN-10249", "Zenith Consulting", "Enterprise Suite", 4, 4500, 18000, "Northeast", "Sarah Chen"],
        ["2025-05-28", "TXN-10250", "Cascade Systems", "Pro License", 15, 850, 12750, "West Coast", "Marcus Johnson"],
        ["2025-06-09", "TXN-10251", "Meridian Corp", "Enterprise Suite", 2, 4500, 9000, "Northeast", "Priya Sharma"],
        ["2025-06-21", "TXN-10252", "Atlas Industries", "Starter Pack", 30, 200, 6000, "West Coast", "David Kim"],
        ["2025-07-03", "TXN-10253", "Pinnacle Group", "Pro License", 10, 850, 8500, "Southeast", "Sarah Chen"],
        ["2025-07-19", "TXN-10254", "Nova Technologies", "Enterprise Suite", 1, 4500, 4500, "Midwest", "Marcus Johnson"],
        ["2025-08-02", "TXN-10255", "Summit Partners", "Pro License", 6, 850, 5100, "Northeast", "Priya Sharma"],
    ]
    for r, row_data in enumerate(details_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_details.cell(row=r, column=c, value=val)
            if c in (6, 7):
                cell.number_format = '$#,##0.00'

    ws_details.column_dimensions["A"].width = 14
    ws_details.column_dimensions["B"].width = 14
    ws_details.column_dimensions["C"].width = 20
    ws_details.column_dimensions["D"].width = 18
    ws_details.column_dimensions["E"].width = 10
    ws_details.column_dimensions["F"].width = 12
    ws_details.column_dimensions["G"].width = 12
    ws_details.column_dimensions["H"].width = 14
    ws_details.column_dimensions["I"].width = 16

    # === Sheet 3: Charts ===
    ws_charts = wb.create_sheet("Charts")

    # Data for chart
    chart_headers = ["Quarter", "Revenue", "Costs", "Profit"]
    for c, h in enumerate(chart_headers, 1):
        ws_charts.cell(row=1, column=c, value=h)
    style_header(ws_charts, 1, len(chart_headers))

    chart_data = [
        ["Q1 2025", 245800, 128400, 117400],
        ["Q2 2025", 312500, 145200, 167300],
        ["Q3 2025", 287600, 139800, 147800],
        ["Q4 2025", 356200, 152600, 203600],
    ]
    for r, row_data in enumerate(chart_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_charts.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '$#,##0'

    # Add a bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Quarterly Financial Performance"
    chart.y_axis.title = "Amount ($)"
    chart.x_axis.title = "Quarter"
    data_ref = Reference(ws_charts, min_col=2, min_row=1, max_col=4, max_row=5)
    cats_ref = Reference(ws_charts, min_col=1, min_row=2, max_row=5)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.width = 20
    chart.height = 12
    ws_charts.add_chart(chart, "A8")

    ws_charts.column_dimensions["A"].width = 14
    ws_charts.column_dimensions["B"].width = 14
    ws_charts.column_dimensions["C"].width = 14
    ws_charts.column_dimensions["D"].width = 14

    # === Sheet 4: Raw Data ===
    ws_raw = wb.create_sheet("Raw Data")

    raw_headers = ["Record ID", "Timestamp", "Source", "Category", "Value", "Status", "Notes"]
    for c, h in enumerate(raw_headers, 1):
        ws_raw.cell(row=1, column=c, value=h)
    style_header(ws_raw, 1, len(raw_headers))

    raw_data = [
        [1001, "2025-01-03 09:14:22", "CRM Import", "Lead Generation", 3420.50, "Verified", "Batch import from HubSpot"],
        [1002, "2025-01-03 09:14:23", "CRM Import", "Customer Support", 1250.00, "Verified", "Support ticket costs"],
        [1003, "2025-01-05 14:30:01", "Manual Entry", "Infrastructure", 8750.00, "Pending Review", "AWS monthly charges"],
        [1004, "2025-01-07 11:22:45", "API Sync", "Lead Generation", 2180.75, "Verified", "Google Ads campaign"],
        [1005, "2025-01-10 08:45:12", "CRM Import", "Sales", 15600.00, "Verified", "Q1 pipeline deals"],
        [1006, "2025-01-12 16:33:08", "Manual Entry", "Marketing", 4320.00, "Verified", "Content creation budget"],
        [1007, "2025-01-15 10:05:30", "API Sync", "Infrastructure", 2100.00, "Pending Review", "Azure services"],
        [1008, "2025-01-18 13:47:55", "CRM Import", "Customer Support", 890.25, "Verified", "Zendesk integration"],
        [1009, "2025-01-20 09:12:18", "Manual Entry", "Sales", 22400.00, "Verified", "Enterprise deal close"],
        [1010, "2025-01-22 15:28:44", "API Sync", "Marketing", 5670.00, "Verified", "Social media campaigns"],
        [1011, "2025-01-25 11:55:02", "CRM Import", "Lead Generation", 1890.50, "Pending Review", "LinkedIn Ads"],
        [1012, "2025-01-28 14:10:33", "Manual Entry", "Infrastructure", 6200.00, "Verified", "Data center costs"],
        [1013, "2025-02-01 08:30:15", "API Sync", "Sales", 18900.00, "Verified", "February pipeline"],
        [1014, "2025-02-03 10:45:22", "CRM Import", "Customer Support", 1560.75, "Verified", "Support staffing"],
        [1015, "2025-02-05 12:20:48", "Manual Entry", "Marketing", 7800.00, "Pending Review", "Event sponsorship"],
    ]
    for r, row_data in enumerate(raw_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_raw.cell(row=r, column=c, value=val)
            if c == 5:
                cell.number_format = '$#,##0.00'

    ws_raw.column_dimensions["A"].width = 12
    ws_raw.column_dimensions["B"].width = 22
    ws_raw.column_dimensions["C"].width = 16
    ws_raw.column_dimensions["D"].width = 20
    ws_raw.column_dimensions["E"].width = 14
    ws_raw.column_dimensions["F"].width = 16
    ws_raw.column_dimensions["G"].width = 28

    # Ensure all sheets are in default state (no special print selection)
    # All sheets should have tabSelected = False except the active one
    wb.active = 0  # Summary is active
    for ws in wb.worksheets:
        ws.views.sheetView[0].tabSelected = False
    # Set only the active sheet as selected (default behavior)
    ws_summary.views.sheetView[0].tabSelected = True

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
