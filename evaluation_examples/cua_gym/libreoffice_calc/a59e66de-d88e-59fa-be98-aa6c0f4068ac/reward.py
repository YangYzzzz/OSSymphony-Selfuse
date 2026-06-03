"""
Reward Script: Build a quota planning model that distributes annual team quota
across reps based on territory potential, historical performance, and ramp status.
Task ID: calc_sales_063
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Weight Factor formulas in E2:E6 (Full=1, Ramping=0.7, New=0.5)
  Component 2 (0.35): Raw Allocation formulas in F2:F6 (weighted blend of territory + revenue * weight)
  Component 3 (0.35): Adjusted Quota formulas in G2:G6 (proportional distribution summing to Team Target)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_063'

# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: QuotaPlan sheet must exist
    if 'QuotaPlan' not in wb.sheetnames:
        print("CRITICAL: 'QuotaPlan' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['QuotaPlan']

    # Ground truth values from context
    # Ramp status: Full=1, Ramping=0.7, New=0.5
    expected_weights = {2: 1, 3: 1, 4: 0.7, 5: 1, 6: 0.5}
    reps = {2: 'Alice', 3: 'Bob', 4: 'Carol', 5: 'Dan', 6: 'Eve'}

    # Precompute expected raw allocation and adjusted quota values
    # B values (Territory Potential)
    b_vals = {2: 5.0, 3: 3.5, 4: 4.0, 5: 2.5, 6: 6.0}
    sum_b = sum(b_vals.values())  # 21.0

    # C values (Last Year Revenue)
    c_vals = {2: 520000, 3: 380000, 4: 150000, 5: 290000, 6: 0}
    sum_c = sum(c_vals.values())  # 1340000

    team_target = 3000000

    # Expected F (raw allocation) = (B/sum_B * 0.5 + C/sum_C * 0.5) * E
    expected_f = {}
    for r in range(2, 7):
        e = expected_weights[r]
        f_val = (b_vals[r] / sum_b * 0.5 + c_vals[r] / sum_c * 0.5) * e
        expected_f[r] = f_val

    sum_f = sum(expected_f.values())

    # Expected G (adjusted quota) = F / sum_F * team_target
    expected_g = {}
    for r in range(2, 7):
        expected_g[r] = expected_f[r] / sum_f * team_target

    # ---------- Component 1: Weight Factor formulas in E2:E6 (0.30 points) ----------
    try:
        e_pass_count = 0
        for row in range(2, 7):
            cell_val = ws.cell(row=row, column=5).value  # Column E
            expected_e = expected_weights[row]

            if cell_val is None:
                print(f"FAIL: E{row} is empty (expected weight factor for {reps[row]})")
                continue

            # Check if it's a formula containing IF logic
            is_formula = isinstance(cell_val, str) and cell_val.startswith('=')

            if is_formula:
                # Verify the formula contains IF and references D column
                formula_upper = cell_val.upper().replace(" ", "")
                if 'IF(' in formula_upper and f'D{row}' in cell_val.upper().replace(" ", ""):
                    print(f"PASS: E{row} has IF formula referencing D{row}: {cell_val}")
                    e_pass_count += 1
                else:
                    print(f"FAIL: E{row} formula doesn't use IF with D{row}: {cell_val}")
            else:
                # Accept numeric value if it matches expected
                try:
                    numeric_val = float(cell_val)
                    if abs(numeric_val - expected_e) < 0.01:
                        print(f"PASS: E{row} has correct numeric value: {numeric_val}")
                        e_pass_count += 1
                    else:
                        print(f"FAIL: E{row} value {numeric_val} != expected {expected_e}")
                except (ValueError, TypeError):
                    print(f"FAIL: E{row} unexpected value: {cell_val}")

        if e_pass_count == 5:
            print(f"PASS: Component 1 — All 5 Weight Factor values correct (0.30 pts)")
            total_score += 0.30
        elif e_pass_count >= 3:
            partial = round(0.30 * e_pass_count / 5, 2)
            print(f"PARTIAL: Component 1 — {e_pass_count}/5 Weight Factor values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {e_pass_count}/5 Weight Factor values correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------- Component 2: Raw Allocation formulas in F2:F6 (0.35 points) ----------
    try:
        f_pass_count = 0
        # Load with data_only to get computed values
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['QuotaPlan']

        for row in range(2, 7):
            formula_cell = ws.cell(row=row, column=6).value  # Column F formula
            data_cell = ws_data.cell(row=row, column=6).value  # Column F computed value

            if formula_cell is None:
                print(f"FAIL: F{row} is empty (expected Raw Allocation for {reps[row]})")
                continue

            is_formula = isinstance(formula_cell, str) and formula_cell.startswith('=')

            if is_formula:
                # Verify formula references B, C, E columns and uses SUM
                formula_upper = formula_cell.upper().replace(" ", "")
                has_b_ref = f'B{row}' in formula_upper or 'B$' in formula_upper
                has_c_ref = f'C{row}' in formula_upper or 'C$' in formula_upper
                has_e_ref = f'E{row}' in formula_upper or 'E$' in formula_upper
                has_sum = 'SUM(' in formula_upper

                if has_b_ref and has_c_ref and has_e_ref and has_sum:
                    print(f"PASS: F{row} has valid allocation formula: {formula_cell}")
                    f_pass_count += 1
                else:
                    print(f"FAIL: F{row} formula missing required refs (B,C,E,SUM): {formula_cell}")
            else:
                # Accept numeric value if close to expected
                try:
                    numeric_val = float(formula_cell)
                    if abs(numeric_val - expected_f[row]) < 0.01:
                        print(f"PASS: F{row} has correct numeric value: {numeric_val}")
                        f_pass_count += 1
                    else:
                        print(f"FAIL: F{row} value {numeric_val} != expected {expected_f[row]:.6f}")
                except (ValueError, TypeError):
                    print(f"FAIL: F{row} unexpected value: {formula_cell}")

        if f_pass_count == 5:
            print(f"PASS: Component 2 — All 5 Raw Allocation formulas correct (0.35 pts)")
            total_score += 0.35
        elif f_pass_count >= 3:
            partial = round(0.35 * f_pass_count / 5, 2)
            print(f"PARTIAL: Component 2 — {f_pass_count}/5 Raw Allocation formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {f_pass_count}/5 Raw Allocation formulas correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------- Component 3: Adjusted Quota formulas in G2:G6 (0.35 points) ----------
    try:
        g_pass_count = 0

        for row in range(2, 7):
            formula_cell = ws.cell(row=row, column=7).value  # Column G formula
            data_cell = ws_data.cell(row=row, column=7).value  # Column G computed value

            if formula_cell is None:
                print(f"FAIL: G{row} is empty (expected Adjusted Quota for {reps[row]})")
                continue

            is_formula = isinstance(formula_cell, str) and formula_cell.startswith('=')

            if is_formula:
                # Verify formula references F column, uses SUM(F), and references I$2 (team target)
                formula_upper = formula_cell.upper().replace(" ", "")
                has_f_ref = f'F{row}' in formula_upper or 'F$' in formula_upper
                has_sum_f = 'SUM(F' in formula_upper
                has_target = 'I' in formula_upper  # references I column (team target)

                if has_f_ref and has_sum_f and has_target:
                    print(f"PASS: G{row} has valid adjusted quota formula: {formula_cell}")
                    g_pass_count += 1
                else:
                    print(f"FAIL: G{row} formula missing required refs (F,SUM(F),I): {formula_cell}")
            else:
                # Accept numeric value if close to expected
                try:
                    numeric_val = float(formula_cell)
                    if abs(numeric_val - expected_g[row]) / expected_g[row] < 0.05:
                        print(f"PASS: G{row} has correct numeric value: {numeric_val:.2f} (expected {expected_g[row]:.2f})")
                        g_pass_count += 1
                    else:
                        print(f"FAIL: G{row} value {numeric_val:.2f} != expected {expected_g[row]:.2f}")
                except (ValueError, TypeError):
                    print(f"FAIL: G{row} unexpected value: {formula_cell}")

        if g_pass_count == 5:
            print(f"PASS: Component 3 — All 5 Adjusted Quota formulas correct (0.35 pts)")
            total_score += 0.35
        elif g_pass_count >= 3:
            partial = round(0.35 * g_pass_count / 5, 2)
            print(f"PARTIAL: Component 3 — {g_pass_count}/5 Adjusted Quota formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {g_pass_count}/5 Adjusted Quota formulas correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
