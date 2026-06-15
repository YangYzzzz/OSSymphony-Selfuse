"""
Initial Setup: Build a quota planning model spreadsheet with rep data.
Task ID: calc_sales_063
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: QuotaPlan ---
    ws = wb.active
    ws.title = 'QuotaPlan'

    # Headers in row 1
    headers = {
        'A1': 'Rep',
        'B1': 'Territory Potential ($M)',
        'C1': 'Last Year Revenue',
        'D1': 'Ramp Status',
        'E1': 'Weight Factor',
        'F1': 'Raw Allocation',
        'G1': 'Adjusted Quota',
    }
    for coord, val in headers.items():
        ws[coord] = val

    # Team target in I1:I2
    ws['I1'] = 'Team Target'
    ws['I2'] = 3000000

    # Rep data rows 2-6
    reps = [
        ('Alice',  5.0, 520000, 'Full'),
        ('Bob',    3.5, 380000, 'Full'),
        ('Carol',  4.0, 150000, 'Ramping'),
        ('Dan',    2.5, 290000, 'Full'),
        ('Eve',    6.0, 0,      'New'),
    ]
    for r, (name, potential, revenue, status) in enumerate(reps, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=potential)
        ws.cell(row=r, column=3, value=revenue)
        ws.cell(row=r, column=4, value=status)
        # E, F, G columns intentionally left empty - task is to build formulas

    # Style headers: bold with light blue background
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    header_font = Font(bold=True, size=11)
    for col in range(1, 10):  # A through I
        cell = ws.cell(row=1, column=col)
        if cell.value is not None:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 4
    ws.column_dimensions['I'].width = 16

    # Format currency columns
    for r in range(2, 7):
        ws.cell(row=r, column=3).number_format = '$#,##0'
    ws['I2'].number_format = '$#,##0'

    # Format territory potential as number with 1 decimal
    for r in range(2, 7):
        ws.cell(row=r, column=2).number_format = '0.0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
