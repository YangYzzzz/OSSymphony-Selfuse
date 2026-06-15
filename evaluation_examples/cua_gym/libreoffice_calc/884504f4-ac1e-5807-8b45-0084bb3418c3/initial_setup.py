"""
Initial Setup: Sheet protection with editable cell range
Task ID: calc_tbl_036
Domain: libreoffice_calc

Creates a budget workbook with Sheet1 protected (password 'budget2024').
All cells locked. B5:B10 contain old monthly target values.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Budget Targets ---
    ws = wb.active
    ws.title = 'Sheet1'

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Title row
    ws.merge_cells('A1:D1')
    ws['A1'] = 'FY2024 Monthly Budget Targets'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Blank row 2
    # Headers in row 3
    headers = ['Month', 'Target ($)', 'Actual ($)', 'Variance ($)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows 4-15 (12 months)
    months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    # Old targets (B5:B10 correspond to rows 5-10 => Feb through Jul)
    old_targets = [750, 820, 880, 910, 960, 1020, 1080, 1150, 1200, 1250, 1300, 1350]
    actuals = [740, 835, 865, 925, 945, 1035, 1095, 1140, 1215, 1260, 1310, 1345]

    for i, month in enumerate(months):
        row = 4 + i
        target = old_targets[i]
        actual = actuals[i]
        variance = actual - target

        ws.cell(row=row, column=1, value=month).border = thin_border
        ws.cell(row=row, column=2, value=target).border = thin_border
        ws.cell(row=row, column=2).number_format = '#,##0'
        ws.cell(row=row, column=3, value=actual).border = thin_border
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=4, value=variance).border = thin_border
        ws.cell(row=row, column=4).number_format = '#,##0'

    # Summary row
    summary_row = 16
    ws.cell(row=summary_row, column=1, value='Total').font = Font(bold=True)
    ws.cell(row=summary_row, column=1).border = thin_border
    ws.cell(row=summary_row, column=2, value=f'=SUM(B4:B15)').border = thin_border
    ws.cell(row=summary_row, column=2).number_format = '#,##0'
    ws.cell(row=summary_row, column=2).font = Font(bold=True)
    ws.cell(row=summary_row, column=3, value=f'=SUM(C4:C15)').border = thin_border
    ws.cell(row=summary_row, column=3).number_format = '#,##0'
    ws.cell(row=summary_row, column=3).font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=f'=SUM(D4:D15)').border = thin_border
    ws.cell(row=summary_row, column=4).number_format = '#,##0'
    ws.cell(row=summary_row, column=4).font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # --- Sheet2: Notes ---
    ws2 = wb.create_sheet('Notes')
    ws2['A1'] = 'Budget Review Notes'
    ws2['A1'].font = Font(size=12, bold=True)
    ws2['A3'] = 'Q1 targets are based on historical averages.'
    ws2['A4'] = 'Q2 targets adjusted for seasonal demand increase.'
    ws2['A5'] = 'Q3-Q4 targets include projected growth of 5%.'
    ws2['A7'] = 'Last updated: 2024-01-10 by Finance Department'

    # Ensure all cells in Sheet1 are locked (default behavior)
    # openpyxl cells are locked by default, but let's be explicit
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=4):
        for cell in row:
            cell.protection = Protection(locked=True)

    # Protect Sheet1 with password
    ws.protection = SheetProtection(
        sheet=True,
        password='budget2024',
        formatCells=False,
        formatColumns=False,
        formatRows=False,
        insertColumns=False,
        insertRows=False,
        insertHyperlinks=False,
        deleteColumns=False,
        deleteRows=False,
        selectLockedCells=False,
        sort=False,
        autoFilter=False,
        pivotTables=False,
        selectUnlockedCells=False,
    )

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
