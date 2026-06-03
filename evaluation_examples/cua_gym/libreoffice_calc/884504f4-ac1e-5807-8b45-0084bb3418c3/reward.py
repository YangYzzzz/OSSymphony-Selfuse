"""
Reward Script: Unprotect sheet, update B5:B10, re-protect with B5:B10 editable
Task ID: calc_tbl_036
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): B5:B10 contain new target values (1000, 1200, 1100, 1300, 1400, 1500)
  Component 2 (0.3): B5:B10 are unlocked (locked=False) within the protected sheet
  Component 3 (0.3): Sheet1 is still protected after re-protection
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_036'

# Expected new values for B5:B10
EXPECTED_VALUES = {
    5: 1000,
    6: 1200,
    7: 1100,
    8: 1300,
    9: 1400,
    10: 1500,
}


def persist_app_state(domain: str):
    """Best-effort save via Ctrl+S in case file is still open in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify Sheet1 exists
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet1']

    # Component 1: B5:B10 contain the correct new values (0.4 points)
    # Initial values are different (820, 880, 910, 960, 1020, 1080),
    # so this only passes on golden.
    try:
        correct_count = 0
        for row, expected in EXPECTED_VALUES.items():
            actual = ws.cell(row=row, column=2).value
            if actual is not None:
                try:
                    if abs(float(actual) - expected) < 0.01:
                        correct_count += 1
                    else:
                        print(f"FAIL: B{row} expected {expected}, found {actual}")
                except (ValueError, TypeError):
                    print(f"FAIL: B{row} value '{actual}' is not numeric")
            else:
                print(f"FAIL: B{row} is empty")

        if correct_count == 6:
            print(f"PASS: Component 1 -- All 6 cells B5:B10 have correct new values (0.4 pts)")
            total_score += 0.4
        elif correct_count >= 4:
            partial = round(0.4 * (correct_count / 6), 2)
            print(f"PARTIAL: Component 1 -- {correct_count}/6 cells correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {correct_count}/6 cells have correct values")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: B5:B10 are unlocked (locked=False) within the sheet (0.3 points)
    # Initially all cells are locked=True, so this only passes on golden.
    try:
        unlocked_count = 0
        for row in range(5, 11):
            cell = ws.cell(row=row, column=2)
            if cell.protection.locked is False:
                unlocked_count += 1
            else:
                print(f"FAIL: B{row} is still locked (expected unlocked)")

        if unlocked_count == 6:
            print(f"PASS: Component 2 -- All 6 cells B5:B10 are unlocked (0.3 pts)")
            total_score += 0.3
        elif unlocked_count >= 4:
            partial = round(0.3 * (unlocked_count / 6), 2)
            print(f"PARTIAL: Component 2 -- {unlocked_count}/6 cells unlocked ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {unlocked_count}/6 cells unlocked")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Sheet1 is still protected (0.3 points)
    # IMPORTANT: Sheet is protected in BOTH initial and golden states.
    # However, this component is combined with Component 2 — we check that
    # the sheet is protected AND B5:B10 are unlocked (the combination is the task change).
    # To avoid scoring a pre-existing condition, we only award these points
    # if at least one B5:B10 cell is also unlocked (i.e., re-protection happened
    # with the cell unlock applied).
    try:
        sheet_protected = ws.protection.sheet is True
        if sheet_protected and unlocked_count > 0:
            print(f"PASS: Component 3 -- Sheet is protected with B5:B10 unlocked (0.3 pts)")
            total_score += 0.3
        elif sheet_protected and unlocked_count == 0:
            print(f"FAIL: Component 3 -- Sheet is protected but no cells were unlocked (pre-existing state)")
        elif not sheet_protected:
            print(f"FAIL: Component 3 -- Sheet is NOT protected (expected re-protected)")
        else:
            print(f"FAIL: Component 3 -- Unexpected state")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook — save in case file is open in LibreOffice GUI
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
