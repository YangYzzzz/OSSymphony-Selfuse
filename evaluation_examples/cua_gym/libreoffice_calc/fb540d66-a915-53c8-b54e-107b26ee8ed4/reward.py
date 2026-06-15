"""
Reward Script: Compound Interest Calculator with Multiple Scenarios
Task ID: calc_wf_024
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Comparison table formulas in C9:C13
  Component 2 (0.25): Growth table populated with values in B18:L22
  Component 3 (0.25): Line chart with 5 series present
  Component 4 (0.25): Currency formatting on dollar value cells
"""

import os
import math
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_024'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Calculator' sheet must exist
    if 'Calculator' not in wb.sheetnames:
        print("CRITICAL: 'Calculator' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Calculator']

    # ---------------------------------------------------------------
    # Component 1: Comparison table formulas in C9:C13 (0.25 points)
    # Initial state: C9:C13 are empty. Golden state has compound interest formulas.
    # We check that each cell has a formula or a numeric value consistent with
    # FV = Principal*(1+Rate/n)^(n*Years).
    # ---------------------------------------------------------------
    try:
        # Expected n values for each row
        freq_n = {9: 1, 10: 2, 11: 4, 12: 12, 13: 365}
        # Expected FV = 10000*(1+0.07/n)^(n*10)
        formula_count = 0
        for row_num, n in freq_n.items():
            cell_val = ws.cell(row=row_num, column=3).value
            if cell_val is not None:
                # Check if it's a formula string containing key elements
                if isinstance(cell_val, str) and '=' in cell_val:
                    formula_count += 1
                elif isinstance(cell_val, (int, float)):
                    # Could be a computed value; check approximate correctness
                    expected_fv = 10000 * (1 + 0.07 / n) ** (n * 10)
                    if abs(float(cell_val) - expected_fv) < 1.0:
                        formula_count += 1
                    else:
                        print(f"FAIL: C{row_num} value {cell_val} != expected {expected_fv:.2f}")
            else:
                print(f"FAIL: C{row_num} is empty")

        if formula_count == 5:
            print(f"PASS: Component 1 - All 5 comparison table formulas/values present (0.25 pts)")
            total_score += 0.25
        elif formula_count >= 3:
            partial = round(0.25 * formula_count / 5, 2)
            print(f"PARTIAL: Component 1 - {formula_count}/5 comparison formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {formula_count}/5 comparison formulas present")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # ---------------------------------------------------------------
    # Component 2: Growth table values in B18:L22 (0.25 points)
    # Initial state: B18:L22 are all empty. Golden has year-by-year values.
    # Check that cells have numeric values (formulas or hardcoded).
    # ---------------------------------------------------------------
    try:
        filled_count = 0
        total_cells = 55  # 5 rows * 11 columns (B through L, rows 18-22)
        correct_count = 0

        freqs = [1, 2, 4, 12, 365]  # n values for rows 18-22
        for row_idx, n in enumerate(freqs):
            row_num = 18 + row_idx
            for col_idx in range(11):  # columns B(2) through L(12), years 0-10
                col_num = 2 + col_idx
                year = col_idx
                cell_val = ws.cell(row=row_num, column=col_num).value
                if cell_val is not None:
                    filled_count += 1
                    # Check value correctness
                    expected_val = 10000 * (1 + 0.07 / n) ** (n * year)
                    try:
                        actual = float(str(cell_val).replace(',', '').replace('$', ''))
                        if abs(actual - expected_val) < 1.0:
                            correct_count += 1
                    except (ValueError, TypeError):
                        pass

        if correct_count >= 50:
            print(f"PASS: Component 2 - Growth table has {correct_count}/{total_cells} correct values (0.25 pts)")
            total_score += 0.25
        elif filled_count >= 30:
            partial = round(0.25 * min(filled_count, total_cells) / total_cells, 2)
            print(f"PARTIAL: Component 2 - {filled_count}/{total_cells} cells filled ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {filled_count}/{total_cells} cells filled in growth table")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # ---------------------------------------------------------------
    # Component 3: Line chart with 5 series (0.25 points)
    # Initial state: 0 charts. Golden has 1 LineChart with 5 series.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Find a line chart
            line_chart = None
            for c in charts:
                if isinstance(c, openpyxl.chart.LineChart):
                    line_chart = c
                    break
            if line_chart is None:
                # Accept any chart type as long as it has series
                line_chart = charts[0]
                print(f"NOTE: Chart is {type(line_chart).__name__}, not LineChart")

            series_count = len(line_chart.series)
            if series_count >= 5:
                print(f"PASS: Component 3 - Chart present with {series_count} series (0.25 pts)")
                total_score += 0.25
            elif series_count >= 3:
                partial = round(0.25 * series_count / 5, 2)
                print(f"PARTIAL: Component 3 - Chart has {series_count}/5 series ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 - Chart has only {series_count} series, expected 5")
        else:
            print(f"FAIL: Component 3 - No charts found (expected line chart with 5 series)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # ---------------------------------------------------------------
    # Component 4: Currency formatting on dollar value cells (0.25 points)
    # Initial state: B2 is 'General', C9:C13 empty, B18:L22 empty.
    # Golden state: B2, C9:C13, B18:L22 all have '$#,##0.00' format.
    # We check that currency format is applied to these cells.
    # ---------------------------------------------------------------
    try:
        currency_cells_found = 0
        currency_cells_total = 0

        # Check B2 (principal)
        fmt_b2 = ws.cell(row=2, column=2).number_format
        if '$' in str(fmt_b2):
            currency_cells_found += 1
        currency_cells_total += 1

        # Check C9:C13 (comparison future values)
        for row_num in range(9, 14):
            cell = ws.cell(row=row_num, column=3)
            if cell.value is not None:
                currency_cells_total += 1
                if '$' in str(cell.number_format):
                    currency_cells_found += 1

        # Check a sample of growth table cells (B18, L18, B22, L22, G20)
        sample_cells = [(18, 2), (18, 12), (22, 2), (22, 12), (20, 7)]
        for r, c in sample_cells:
            cell = ws.cell(row=r, column=c)
            if cell.value is not None:
                currency_cells_total += 1
                if '$' in str(cell.number_format):
                    currency_cells_found += 1

        if currency_cells_total > 0 and currency_cells_found >= currency_cells_total * 0.8:
            print(f"PASS: Component 4 - Currency formatting on {currency_cells_found}/{currency_cells_total} checked cells (0.25 pts)")
            total_score += 0.25
        elif currency_cells_found >= 3:
            partial = round(0.25 * currency_cells_found / max(currency_cells_total, 1), 2)
            print(f"PARTIAL: Component 4 - Currency formatting on {currency_cells_found}/{currency_cells_total} cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - Currency formatting on only {currency_cells_found}/{currency_cells_total} cells")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
