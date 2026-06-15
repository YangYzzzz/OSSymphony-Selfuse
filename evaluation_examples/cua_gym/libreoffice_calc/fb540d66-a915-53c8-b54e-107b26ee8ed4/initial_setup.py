"""
Initial Setup: Compound interest calculator with multiple scenarios
Task ID: calc_wf_024
Domain: libreoffice_calc

Creates the pre-task state: input section, empty comparison table structure,
and empty growth curve table structure. NO formulas, NO computed values,
NO charts, NO currency formatting.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calculator"

    # ── Input Section (A1:B5) ──
    header_font = Font(name="Calibri", size=12, bold=True)
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_font_white = Font(name="Calibri", size=12, bold=True, color="FFFFFF")

    ws["A1"] = "Parameter"
    ws["B1"] = "Value"
    for cell in [ws["A1"], ws["B1"]]:
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    input_data = [
        ("Principal", 10000),
        ("Annual Rate", 0.07),
        ("Compounding Frequency", ""),
        ("Years", 10),
    ]
    for r, (label, val) in enumerate(input_data, 2):
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=11)
        ws.cell(row=r, column=2, value=val).font = Font(name="Calibri", size=11)

    # Rate as percentage display (just the number, no currency format yet)
    ws["B3"].number_format = '0%'

    # ── Comparison Table (rows 7-13) ──
    ws["A7"] = "Compound Interest Comparison"
    ws["A7"].font = Font(name="Calibri", size=13, bold=True)

    comp_headers = ["Compounding Frequency", "n (times/year)", "Future Value"]
    for c, h in enumerate(comp_headers, 1):
        cell = ws.cell(row=8, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    frequencies = [
        ("Annual", 1),
        ("Semi-Annual", 2),
        ("Quarterly", 4),
        ("Monthly", 12),
        ("Daily", 365),
    ]
    for r, (name, n) in enumerate(frequencies, 9):
        ws.cell(row=r, column=1, value=name).font = Font(name="Calibri", size=11)
        ws.cell(row=r, column=2, value=n).font = Font(name="Calibri", size=11)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")
        # Column C (Future Value) left EMPTY — task is to fill these

    # ── Growth Curve Table (rows 16+) ──
    ws["A16"] = "Growth Over Time by Compounding Frequency"
    ws["A16"].font = Font(name="Calibri", size=13, bold=True)

    # Row 17: headers — "Year" then 0..10
    ws.cell(row=17, column=1, value="Year").font = header_font_white
    ws.cell(row=17, column=1).fill = header_fill
    ws.cell(row=17, column=1).alignment = Alignment(horizontal="center")
    for yr in range(11):
        cell = ws.cell(row=17, column=yr + 2, value=yr)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Rows 18-22: frequency labels only (values left EMPTY)
    for r, (name, _n) in enumerate(frequencies, 18):
        cell = ws.cell(row=r, column=1, value=name)
        cell.font = Font(name="Calibri", size=11, bold=True)

    # ── Column widths ──
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18
    for col_letter_idx in range(2, 13):  # columns B-L for growth table
        from openpyxl.utils import get_column_letter
        letter = get_column_letter(col_letter_idx)
        if ws.column_dimensions[letter].width < 14:
            ws.column_dimensions[letter].width = 14

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
