"""
Reward Script: Set up batch record tracking with a sheet per product line.
Task ID: calc_ops_qc_batch_records_023
Domain: libreoffice_calc

Scoring rubric:
  Component 1 (0.4): Three line sheets (Line-A, Line-B, Line-C) exist with correct
                     8-column batch tracking headers matching the Template structure.
  Component 2 (0.3): Summary sheet exists with correct 5-column headers:
                     Line, Total Batches, Total Produced, Total Rejected, Overall Reject Rate.
  Component 3 (0.3): Summary rows 2-4 have the correct line names and reference formulas
                     using COUNTA for batch count and SUM for quantity columns per line.
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_qc_batch_records_023'

EXPECTED_HEADERS = [
    'Batch Number', 'Production Date', 'Shift', 'Operator',
    'Qty Produced', 'Qty Rejected', 'Reject Rate %', 'Status'
]
EXPECTED_LINE_SHEETS = ['Line-A', 'Line-B', 'Line-C']
EXPECTED_SUMMARY_HEADERS = [
    'Line', 'Total Batches', 'Total Produced', 'Total Rejected', 'Overall Reject Rate'
]
EXPECTED_LINE_NAMES = ['Line-A', 'Line-B', 'Line-C']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Component 1: Line-A, Line-B, Line-C sheets exist with correct 8-column headers (0.4 points)
    # Each of the 3 sheets must exist AND contain the same headers as the Template.
    # 0.4 points total: each sheet contributes ~0.133 — but we award the whole 0.4 only if ALL 3 are correct.
    try:
        sheets_with_correct_headers = 0
        for sheet_name in EXPECTED_LINE_SHEETS:
            if sheet_name not in sheet_names:
                print(f"FAIL: Component 1 — sheet '{sheet_name}' not found (found: {sheet_names})")
                continue
            ws = wb[sheet_name]
            # Read actual row-1 headers
            actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 9)]
            if actual_headers == EXPECTED_HEADERS:
                print(f"PASS: Component 1 — '{sheet_name}' has correct 8-column headers")
                sheets_with_correct_headers += 1
            else:
                print(f"FAIL: Component 1 — '{sheet_name}' headers mismatch: expected {EXPECTED_HEADERS}, found {actual_headers}")

        if sheets_with_correct_headers == 3:
            print(f"PASS: Component 1 — All 3 line sheets have correct headers (0.4 pts)")
            total_score += 0.4
        elif sheets_with_correct_headers > 0:
            partial = round(sheets_with_correct_headers * 0.4 / 3, 4)
            print(f"PARTIAL: Component 1 — {sheets_with_correct_headers}/3 line sheets correct (+{partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 — No line sheets with correct headers found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Summary sheet exists with correct 5-column headers (0.3 points)
    try:
        if 'Summary' not in sheet_names:
            print(f"FAIL: Component 2 — 'Summary' sheet not found (found: {sheet_names})")
        else:
            ws_sum = wb['Summary']
            actual_sum_headers = [ws_sum.cell(row=1, column=c).value for c in range(1, 6)]
            if actual_sum_headers == EXPECTED_SUMMARY_HEADERS:
                print(f"PASS: Component 2 — Summary sheet has correct 5-column headers (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Summary headers mismatch: expected {EXPECTED_SUMMARY_HEADERS}, found {actual_sum_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary rows 2-4 have correct line names and COUNTA/SUM reference formulas (0.3 points)
    # Checks: column A has line names, columns B-D have formulas referencing each line sheet,
    # column E has a formula for overall reject rate.
    try:
        if 'Summary' not in sheet_names:
            print("FAIL: Component 3 — 'Summary' sheet missing, cannot verify formulas")
        else:
            ws_sum = wb['Summary']
            formula_checks_passed = 0
            for row_idx, line_name in enumerate(EXPECTED_LINE_NAMES, start=2):
                # Check A column has line name
                a_val = ws_sum.cell(row=row_idx, column=1).value
                # Check B column uses COUNTA referencing the correct line sheet
                b_val = ws_sum.cell(row=row_idx, column=2).value
                # Check C column uses SUM referencing E col of line sheet
                c_val = ws_sum.cell(row=row_idx, column=3).value
                # Check D column uses SUM referencing F col of line sheet
                d_val = ws_sum.cell(row=row_idx, column=4).value
                # Check E column has some formula for overall reject rate
                e_val = ws_sum.cell(row=row_idx, column=5).value

                line_ok = (a_val == line_name)
                counta_ok = (
                    isinstance(b_val, str) and
                    'COUNTA' in b_val.upper() and
                    line_name.replace('-', '-') in b_val
                )
                sum_prod_ok = (
                    isinstance(c_val, str) and
                    'SUM' in c_val.upper() and
                    line_name.replace('-', '-') in c_val
                )
                sum_rej_ok = (
                    isinstance(d_val, str) and
                    'SUM' in d_val.upper() and
                    line_name.replace('-', '-') in d_val
                )
                rate_ok = (
                    isinstance(e_val, str) and
                    len(e_val) > 1  # some formula referencing D/C columns
                )

                if line_ok and counta_ok and sum_prod_ok and sum_rej_ok and rate_ok:
                    print(f"PASS: Component 3 — Row {row_idx} '{line_name}': line name correct, COUNTA/SUM/rate formulas found")
                    formula_checks_passed += 1
                else:
                    issues = []
                    if not line_ok:
                        issues.append(f"A{row_idx}={repr(a_val)} (expected {repr(line_name)})")
                    if not counta_ok:
                        issues.append(f"B{row_idx}={repr(b_val)} (expected COUNTA referencing {line_name})")
                    if not sum_prod_ok:
                        issues.append(f"C{row_idx}={repr(c_val)} (expected SUM of E col for {line_name})")
                    if not sum_rej_ok:
                        issues.append(f"D{row_idx}={repr(d_val)} (expected SUM of F col for {line_name})")
                    if not rate_ok:
                        issues.append(f"E{row_idx}={repr(e_val)} (expected reject rate formula)")
                    print(f"FAIL: Component 3 — Row {row_idx} '{line_name}': {'; '.join(issues)}")

            if formula_checks_passed == 3:
                print(f"PASS: Component 3 — All 3 Summary rows have correct formulas (0.3 pts)")
                total_score += 0.3
            elif formula_checks_passed > 0:
                partial = round(formula_checks_passed * 0.3 / 3, 4)
                print(f"PARTIAL: Component 3 — {formula_checks_passed}/3 Summary rows correct (+{partial} pts)")
                total_score += partial
            else:
                print("FAIL: Component 3 — No Summary rows have correct formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
