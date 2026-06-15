"""
Initial Setup: Batch Record Tracking - Template Sheet Only
Task ID: calc_ops_qc_batch_records_023
Domain: libreoffice_calc

Creates a workbook with a single 'Template' sheet containing batch tracking
headers (Row 1 formatted with blue background and white bold text) and
empty data rows (Rows 2-51). No Line-A/B/C or Summary sheets present yet.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_qc_batch_records_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Template Sheet ---
    ws = wb.active
    ws.title = 'Template'

    # Headers
    headers = [
        'Batch Number', 'Production Date', 'Shift', 'Operator',
        'Qty Produced', 'Qty Rejected', 'Reject Rate %', 'Status'
    ]

    # Blue background (#4472C4) + white bold text for header row
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')  # white text

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Set column widths for readability
    col_widths = [15, 16, 10, 18, 14, 14, 14, 12]
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    # Row 1 height
    ws.row_dimensions[1].height = 22

    # Rows 2-51: Empty data entry rows (no data — agent will fill these)
    # Just ensure the rows exist structurally (they are empty by default)
    # We set row height to indicate usable rows
    for row in range(2, 52):
        ws.row_dimensions[row].height = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Template')
    print('Row 1: Headers with blue background and white bold text')
    print('Rows 2-51: Empty data entry rows')


create_initial()
