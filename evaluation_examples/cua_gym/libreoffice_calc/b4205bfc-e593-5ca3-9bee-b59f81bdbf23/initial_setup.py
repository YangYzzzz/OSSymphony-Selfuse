"""
Initial Setup: Employee Onboarding Excel + PDF Template
Task ID: osworld_multi_apps_excel_pdf_form_011
Domain: libreoffice_calc (multi-app: Excel + PDF)

Creates:
  - /home/user/new_hires.xlsx  (employee data spreadsheet)
  - /home/user/Desktop/onboarding_template.pdf  (PDF form template)

Does NOT create:
  - Desktop/onboarding/ folder (agent must create this)
  - Per-employee PDFs (agent must generate these)
"""

import os
import shlex
import subprocess
import time
import sys

# Ensure dependencies are installed on the VM
subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl", "fpdf2", "--quiet"], check=False)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_excel_pdf_form_011'
XLSX_OUTPUT = f'{WORKDIR}/new_hires.xlsx'
DESKTOP = f'{WORKDIR}/Desktop'
PDF_TEMPLATE = f'{DESKTOP}/onboarding_template.pdf'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_xlsx():
    """Create the new_hires.xlsx with realistic employee data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "New Hires"

    # Header row styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        "EmployeeName",
        "EmployeeID",
        "Department",
        "Role",
        "StartDate",
        "Manager",
        "IT Access Level",
        "Equipment Assigned"
    ]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Realistic employee data
    # IT Access Level: Standard / Elevated / Admin
    # Equipment Assigned: Laptop / Desktop / Both
    data = [
        ["Sarah Chen",       "EMP-2025-001", "Engineering",       "Software Engineer",         "2025-03-03", "David Kim",       "Standard",  "Laptop"],
        ["Marcus Johnson",   "EMP-2025-002", "Marketing",         "Digital Marketing Manager", "2025-03-03", "Emily Watson",    "Standard",  "Both"],
        ["Priya Patel",      "EMP-2025-003", "Finance",           "Financial Analyst",         "2025-03-10", "Robert Hughes",   "Elevated",  "Desktop"],
        ["James O'Brien",    "EMP-2025-004", "Engineering",       "DevOps Engineer",           "2025-03-10", "David Kim",       "Admin",     "Laptop"],
        ["Anika Müller",     "EMP-2025-005", "Human Resources",   "HR Business Partner",       "2025-03-17", "Sandra Torres",   "Standard",  "Desktop"],
        ["Chen Wei",         "EMP-2025-006", "Security",          "Information Security Analyst","2025-03-17","Laura Fernandez", "Admin",    "Both"],
        ["Fatima Al-Rashid", "EMP-2025-007", "Product",           "Product Manager",           "2025-03-24", "Michael Chang",   "Elevated",  "Laptop"],
        ["Diego Ramirez",    "EMP-2025-008", "Engineering",       "Backend Developer",         "2025-03-24", "David Kim",       "Standard",  "Laptop"],
        ["Yuki Tanaka",      "EMP-2025-009", "Data Science",      "Machine Learning Engineer", "2025-03-31", "Laura Fernandez", "Elevated",  "Both"],
        ["Olivia Bennett",   "EMP-2025-010", "Legal",             "Corporate Counsel",         "2025-03-31", "Richard Nolan",   "Standard",  "Desktop"],
    ]

    data_align = Alignment(horizontal="left", vertical="center")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = data_align
            cell.border = data_border
        # Alternating row background
        if r % 2 == 0:
            row_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        else:
            row_fill = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = row_fill

    # Column widths
    col_widths = [22, 16, 20, 30, 14, 20, 18, 20]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    wb.save(XLSX_OUTPUT)
    print(f'Initial file created: {XLSX_OUTPUT}')


def create_pdf_template():
    """Create the onboarding_template.pdf form."""
    try:
        from fpdf import FPDF
    except ImportError:
        subprocess.run(["pip3", "install", "fpdf2"], check=True)
        from fpdf import FPDF

    os.makedirs(DESKTOP, exist_ok=True)

    pdf = FPDF()
    pdf.set_margins(20, 20, 20)
    pdf.add_page()

    # ---- Title ----
    pdf.set_font("Helvetica", "B", 18)
    pdf.set_fill_color(46, 117, 182)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 12, "EMPLOYEE ONBOARDING PACKET", border=0, ln=True, align="C", fill=True)
    pdf.ln(4)

    pdf.set_text_color(0, 0, 0)

    # ---- Section: Personal Details ----
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(217, 225, 242)
    pdf.cell(0, 8, "  PERSONAL DETAILS", border=1, ln=True, fill=True)
    pdf.ln(2)

    pdf.set_font("Helvetica", "", 10)
    fields_left = [
        ("Employee Name:", ""),
        ("Employee ID:", ""),
        ("Department:", ""),
        ("Role / Title:", ""),
    ]
    fields_right = [
        ("Start Date:", ""),
        ("Manager:", ""),
        ("Office Location:", ""),
        ("Work Email:", ""),
    ]

    col_w = 85
    label_w = 38
    val_w = col_w - label_w

    for (lbl_l, _), (lbl_r, _) in zip(fields_left, fields_right):
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(label_w, 7, lbl_l, border="B")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(val_w, 7, "", border="B")
        pdf.cell(10, 7, "")  # gap
        pdf.set_font("Helvetica", "B", 10)
        pdf.cell(label_w, 7, lbl_r, border="B")
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(val_w, 7, "", border="B", ln=True)
        pdf.ln(1)

    pdf.ln(4)

    # ---- Section: IT Access Level ----
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(217, 225, 242)
    pdf.cell(0, 8, "  IT ACCESS LEVEL", border=1, ln=True, fill=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "", 11)
    it_levels = ["Standard", "Elevated", "Admin"]
    for level in it_levels:
        pdf.set_font("Helvetica", "", 11)
        # Checkbox placeholder
        pdf.cell(8, 8, "", border=1)   # empty checkbox
        pdf.cell(5, 8, "")
        pdf.cell(60, 8, level)
        pdf.ln(6)

    pdf.ln(2)

    # ---- Section: Equipment Assigned ----
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(217, 225, 242)
    pdf.cell(0, 8, "  EQUIPMENT ASSIGNED", border=1, ln=True, fill=True)
    pdf.ln(3)

    equipment_types = [
        ("Laptop",  "Portable workstation for mobile/remote work"),
        ("Desktop", "Fixed workstation for in-office use"),
        ("Both",    "Both laptop and desktop workstations provided"),
    ]
    for eq_type, eq_desc in equipment_types:
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(8, 8, "", border=1)   # empty checkbox
        pdf.cell(5, 8, "")
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(30, 8, eq_type)
        pdf.set_font("Helvetica", "", 10)
        pdf.cell(0, 8, f"- {eq_desc}", ln=True)
        pdf.ln(1)

    pdf.ln(4)

    # ---- Section: Badge Photo ----
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(217, 225, 242)
    pdf.cell(0, 8, "  BADGE PHOTO", border=1, ln=True, fill=True)
    pdf.ln(3)

    # Photo placeholder box
    pdf.set_fill_color(240, 240, 240)
    pdf.rect(20, pdf.get_y(), 40, 50, "DF")
    x_after = 20 + 40 + 8
    y_photo = pdf.get_y()
    pdf.set_xy(x_after, y_photo + 5)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.multi_cell(
        0, 6,
        "Attach employee badge photo here.\n"
        "Photo will be used for ID card and\n"
        "building access systems.\n\n"
        "Required: 2x2 inch, plain background.",
        border=0
    )
    pdf.set_text_color(0, 0, 0)
    pdf.ln(30)  # space after photo area

    # ---- Section: Acknowledgment ----
    pdf.ln(4)
    pdf.set_font("Helvetica", "B", 12)
    pdf.set_fill_color(217, 225, 242)
    pdf.cell(0, 8, "  ACKNOWLEDGMENT", border=1, ln=True, fill=True)
    pdf.ln(3)

    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(
        0, 5,
        "I acknowledge receipt of this onboarding packet and confirm that the information above is accurate. "
        "I understand the IT access level and equipment assigned to me and agree to use company resources "
        "in accordance with the Acceptable Use Policy.",
        border=0
    )
    pdf.ln(6)

    pdf.set_font("Helvetica", "", 10)
    pdf.cell(70, 7, "Employee Signature:", border="B")
    pdf.cell(20, 7, "")
    pdf.cell(40, 7, "Date:", border="B")
    pdf.cell(0, 7, "", ln=True)
    pdf.ln(6)
    pdf.cell(70, 7, "HR Representative:", border="B")
    pdf.cell(20, 7, "")
    pdf.cell(40, 7, "Date:", border="B")

    pdf.output(PDF_TEMPLATE)
    print(f'PDF template created: {PDF_TEMPLATE}')


def create_initial():
    create_xlsx()
    create_pdf_template()

    # GUI-ready startup: open the xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{XLSX_OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with new_hires.xlsx (DISPLAY=:0)')


create_initial()
