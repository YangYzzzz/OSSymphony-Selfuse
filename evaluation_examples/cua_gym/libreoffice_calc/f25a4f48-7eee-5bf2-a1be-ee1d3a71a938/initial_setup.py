"""
Initial Setup: Create Employee_Performance spreadsheet with 69 employee records.
Task ID: calc_gcv_023
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# Realistic data pools
FIRST_NAMES = [
    "Sarah", "Marcus", "Priya", "James", "Fatima", "Chen", "Olivia", "Raj",
    "Elena", "David", "Aisha", "Kevin", "Yuki", "Andre", "Maria", "Thomas",
    "Nadia", "Robert", "Mei", "Carlos", "Zara", "Patrick", "Sonia", "Michael",
    "Amara", "Daniel", "Hannah", "Leo", "Grace", "Samuel", "Lina", "Owen",
    "Jasmine", "Victor", "Natalie", "Ibrahim", "Chloe", "Ryan", "Deepa", "Eric",
    "Valentina", "Nathan", "Simone", "Alexander", "Rosa", "William", "Anya",
    "Brandon", "Leila", "Scott", "Tanya", "George", "Ingrid", "Felix", "Diana",
    "Hugo", "Mia", "Cedric", "Freya", "Tobias", "Julia", "Dante", "Kira",
    "Ruben", "Sofia", "Neil", "Hana", "Marco", "Lily",
]

LAST_NAMES = [
    "Chen", "Johnson", "Patel", "Williams", "Ahmed", "Wei", "Rodriguez",
    "Sharma", "Thompson", "Kowalski", "Nakamura", "Garcia", "O'Brien",
    "Mueller", "Kim", "Santos", "Larsson", "Okafor", "Brown", "Gupta",
    "Martinez", "Taylor", "Singh", "Anderson", "Ibrahim", "Novak",
    "Fischer", "Morales", "Park", "Davis", "Hernandez", "Sato", "Clark",
    "Petrov", "Ali", "Torres", "Berg", "Wilson", "Reeves", "Hoffman",
    "Grant", "Foster", "Barnes", "Hughes", "Price", "Reed", "Cooper",
    "Phillips", "Mitchell", "Powell", "Long", "Russell", "Hayes", "Cole",
    "Murray", "Ross", "Bell", "Wood", "Ward", "Watson", "Brooks",
    "Sullivan", "Morgan", "Kelly", "Howard", "Cox", "Diaz", "Chapman",
    "Fox",
]

DEPARTMENTS = [
    "Engineering", "Marketing", "Sales", "Finance", "Human Resources",
    "Operations", "Product", "Customer Success", "Legal", "Data Science",
]

MANAGERS = [
    "Jennifer Walsh", "Richard Torres", "Samantha Lee", "Michael Grant",
    "Patricia Hoffman", "David Nakamura", "Laura Reeves", "Robert Singh",
    "Angela Foster", "Christopher Park",
]


def create_initial():
    random.seed(42)  # reproducible
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee_Performance"

    # Headers
    headers = [
        "Emp ID", "Name", "Department", "Manager",
        "Start Date", "Projects Completed", "Client Rating", "Performance Index",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Generate 69 employee rows
    used_names = set()
    for i in range(69):
        row = i + 2
        emp_id = f"EMP-{1001 + i}"

        # Unique name
        while True:
            fn = random.choice(FIRST_NAMES)
            ln = random.choice(LAST_NAMES)
            full = f"{fn} {ln}"
            if full not in used_names:
                used_names.add(full)
                break

        dept = random.choice(DEPARTMENTS)
        mgr = random.choice(MANAGERS)

        # Start date between 2019 and 2025
        year = random.randint(2019, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        start_date = f"{year}-{month:02d}-{day:02d}"

        projects = random.randint(2, 35)
        client_rating = round(random.uniform(2.5, 5.0), 1)
        perf_index = round(random.uniform(0.0, 10.0), 1)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=full)
        ws.cell(row=row, column=3, value=dept)
        ws.cell(row=row, column=4, value=mgr)
        ws.cell(row=row, column=5, value=start_date)
        ws.cell(row=row, column=6, value=projects)
        ws.cell(row=row, column=7, value=client_rating)
        ws.cell(row=row, column=8, value=perf_index)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 18

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
