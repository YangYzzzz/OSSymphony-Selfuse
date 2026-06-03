"""
Initial Setup: Mark margin formula cells as hidden+locked and protect sheet
Task ID: calc_ps_027
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Protection, Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pricing'

    # Headers
    headers = ['Product', 'Cost', 'Markup', 'Price', 'Tax', 'Margin']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Product data (29 rows: rows 2-30)
    products = [
        ('Wireless Mouse', 12.50, 0.45),
        ('USB-C Hub 7-Port', 22.00, 0.55),
        ('Mechanical Keyboard', 35.00, 0.40),
        ('27" Monitor Stand', 18.75, 0.60),
        ('Laptop Sleeve 15"', 8.00, 0.70),
        ('HDMI Cable 6ft', 3.50, 0.80),
        ('Webcam HD 1080p', 28.00, 0.50),
        ('Desk Lamp LED', 15.00, 0.65),
        ('Mouse Pad XL', 6.00, 0.75),
        ('USB Flash Drive 64GB', 5.50, 0.85),
        ('Bluetooth Speaker', 19.00, 0.55),
        ('Phone Stand Adjustable', 7.25, 0.90),
        ('Cable Management Kit', 9.00, 0.60),
        ('Screen Protector 15"', 4.00, 1.00),
        ('Wireless Charger Pad', 11.00, 0.65),
        ('Ergonomic Wrist Rest', 10.50, 0.50),
        ('Portable SSD 500GB', 42.00, 0.35),
        ('Noise Cancelling Earbuds', 55.00, 0.45),
        ('Laptop Cooling Pad', 16.00, 0.55),
        ('Mini Projector', 85.00, 0.30),
        ('Digital Drawing Tablet', 48.00, 0.40),
        ('USB Microphone', 32.00, 0.50),
        ('Smart Power Strip', 14.00, 0.60),
        ('Document Scanner', 65.00, 0.35),
        ('Webcam Ring Light', 12.00, 0.70),
        ('Ethernet Adapter USB-C', 9.50, 0.65),
        ('Portable Monitor 15.6"', 120.00, 0.25),
        ('Wireless Presenter', 18.00, 0.55),
        ('Anti-Glare Screen Filter', 22.50, 0.40),
    ]

    cost_fmt = '$#,##0.00'
    pct_fmt = '0%'
    price_fmt = '$#,##0.00'

    for i, (product, cost, markup) in enumerate(products, 2):
        price = round(cost * (1 + markup), 2)
        tax = round(price * 0.08, 2)

        ws.cell(row=i, column=1, value=product).border = thin_border
        c_cost = ws.cell(row=i, column=2, value=cost)
        c_cost.number_format = cost_fmt
        c_cost.border = thin_border

        c_markup = ws.cell(row=i, column=3, value=markup)
        c_markup.number_format = pct_fmt
        c_markup.border = thin_border

        c_price = ws.cell(row=i, column=4, value=price)
        c_price.number_format = price_fmt
        c_price.border = thin_border

        c_tax = ws.cell(row=i, column=5, value=tax)
        c_tax.number_format = price_fmt
        c_tax.border = thin_border

        # F column: margin formula = (D-B)/D = profit margin percentage
        c_margin = ws.cell(row=i, column=6, value=f'=(D{i}-B{i})/D{i}')
        c_margin.number_format = '0.00%'
        c_margin.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    # All cells are locked by default in openpyxl (Protection(locked=True))
    # None are hidden - this is the default state
    # Sheet is NOT protected - this is the default state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
