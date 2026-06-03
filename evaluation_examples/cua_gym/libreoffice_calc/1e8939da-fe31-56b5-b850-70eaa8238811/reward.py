"""
Reward Script: Mark F2:F30 as Hidden+Locked and protect sheet with password 'price2024'
Task ID: calc_ps_027
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): F2:F30 cells have hidden=True
  Component 2 (0.25): Sheet protection is enabled
  Component 3 (0.25): Sheet password hash matches 'price2024'
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_027'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Pricing' sheet exists (precondition gate)
    if 'Pricing' not in wb.sheetnames:
        print("CRITICAL: 'Pricing' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pricing']

    # Component 1: F2:F30 cells have hidden=True (0.50 points)
    # In initial_env, all cells have hidden=False. In golden_env, F2:F30 should be hidden=True.
    try:
        hidden_count = 0
        total_cells = 29  # F2 through F30
        for r in range(2, 31):
            cell = ws.cell(row=r, column=6)
            if cell.protection.hidden:
                hidden_count += 1

        if hidden_count == total_cells:
            print(f"PASS: Component 1 — All {total_cells} cells F2:F30 have hidden=True (0.50 pts)")
            total_score += 0.50
        elif hidden_count > 0:
            # Partial credit proportional to how many cells are hidden
            partial = 0.50 * (hidden_count / total_cells)
            print(f"PARTIAL: Component 1 — {hidden_count}/{total_cells} cells hidden ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in F2:F30 have hidden=True (0/{total_cells})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sheet protection is enabled (0.25 points)
    # In initial_env, sheet is unprotected. In golden_env, sheet must be protected.
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 2 — Sheet protection is enabled (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Sheet protection is NOT enabled")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet password matches hash for 'price2024' (0.25 points)
    # openpyxl stores the password as a hash. The hash for 'price2024' is '8832'.
    # In initial_env, password is None. In golden_env, it should be '8832'.
    try:
        pwd_hash = ws.protection.password
        # Compute expected hash using openpyxl's own mechanism
        from openpyxl.worksheet.protection import SheetProtection
        expected_sp = SheetProtection(password='price2024', sheet=True)
        expected_hash = expected_sp.password

        if pwd_hash is not None and str(pwd_hash) == str(expected_hash):
            print(f"PASS: Component 3 — Password hash matches 'price2024' (hash={pwd_hash}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Password hash mismatch: expected {expected_hash}, found {pwd_hash}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
