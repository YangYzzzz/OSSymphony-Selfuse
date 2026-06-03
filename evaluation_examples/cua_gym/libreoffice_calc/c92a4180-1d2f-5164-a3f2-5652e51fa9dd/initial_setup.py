"""
Initial Setup: Logistics Dashboard - Multi Pivot Table Task
Task ID: osworld_calc_pivot_multi_styled_009
Domain: libreoffice_calc

Creates a logistics shipment workbook with:
- Sheet1: Raw logistics data (Shipment ID, Carrier, Region, Shipment Type, Delivery Time, Freight Cost)
- Sheet2: Empty sheet (agent will create pivot tables here)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Shipments ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Shipment ID', 'Carrier', 'Region', 'Shipment Type', 'Delivery Time (days)', 'Freight Cost']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")

    # Realistic logistics data
    data = [
        ['SHP-001', 'FedEx',   'North',  'Express',   2,  245.50],
        ['SHP-002', 'UPS',     'South',  'Standard',  5,  132.75],
        ['SHP-003', 'DHL',     'East',   'Express',   3,  310.00],
        ['SHP-004', 'FedEx',   'West',   'Freight',   7,  875.20],
        ['SHP-005', 'USPS',    'North',  'Standard',  4,   89.40],
        ['SHP-006', 'UPS',     'East',   'Express',   2,  298.60],
        ['SHP-007', 'DHL',     'South',  'Freight',   8,  940.00],
        ['SHP-008', 'FedEx',   'West',   'Standard',  5,  154.30],
        ['SHP-009', 'USPS',    'East',   'Express',   3,  201.90],
        ['SHP-010', 'UPS',     'North',  'Freight',   6,  720.45],
        ['SHP-011', 'DHL',     'West',   'Standard',  4,  167.80],
        ['SHP-012', 'FedEx',   'South',  'Express',   2,  275.15],
        ['SHP-013', 'USPS',    'North',  'Standard',  5,   95.60],
        ['SHP-014', 'UPS',     'East',   'Freight',   9, 1050.00],
        ['SHP-015', 'DHL',     'North',  'Express',   3,  320.75],
        ['SHP-016', 'FedEx',   'East',   'Standard',  4,  143.20],
        ['SHP-017', 'USPS',    'West',   'Express',   3,  210.40],
        ['SHP-018', 'UPS',     'South',  'Standard',  5,  128.90],
        ['SHP-019', 'DHL',     'East',   'Freight',   7,  880.00],
        ['SHP-020', 'FedEx',   'North',  'Express',   2,  260.35],
        ['SHP-021', 'USPS',    'South',  'Freight',   8,  815.00],
        ['SHP-022', 'UPS',     'West',   'Express',   2,  290.70],
        ['SHP-023', 'DHL',     'North',  'Standard',  6,  175.50],
        ['SHP-024', 'FedEx',   'South',  'Freight',   7,  920.00],
        ['SHP-025', 'USPS',    'East',   'Standard',  4,   98.25],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Column widths for Sheet1
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 22
    ws1.column_dimensions['F'].width = 14

    # --- Sheet2: Empty (agent will create pivot tables here) ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2 is intentionally empty - agent must create the pivot tables

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
