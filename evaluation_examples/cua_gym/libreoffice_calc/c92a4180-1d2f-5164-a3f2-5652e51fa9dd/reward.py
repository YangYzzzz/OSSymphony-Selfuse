"""
Reward Script: Logistics dashboard - create pivot tables in Sheet2 with styled header
Task ID: osworld_calc_pivot_multi_styled_009
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Merged header row in Sheet2 (A1:F1) with blue fill and bold white text (0.25 pts)
  Component 2: Pivot table 1 - Total Shipments by Carrier (rows 3-8 area) (0.25 pts)
  Component 3: Pivot table 2 - Average Delivery Time by Region (rows 10-15 area) (0.25 pts)
  Component 4: Pivot table 3 - Total Freight Cost by Shipment Type (rows 17-21 area) (0.25 pts)
  Total: 1.0

Context verification values (from task context):
  - Header: 'Logistics Performance Report' in A1, merged A1:F1
  - Header fill: FF4472C4 (blue), font bold, font color white
  - Pivot 1 carriers: DHL=6, FedEx=7, UPS=6, USPS=6
  - Pivot 2 regions: East~4.57, North~4.0, South~5.83, West~4.2
  - Pivot 3 types: Express~2413.35, Freight~6200.65, Standard~1185.70
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_009'


def persist_app_state():
    """Attempt to save any open LibreOffice file before verification."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        import time
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.5)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Create three pivot tables in Sheet2 plus a merged styled header.
    """
    total_score = 0.0

    # Load workbook — if this fails, the file is invalid/missing
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in workbook")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet2']

    # -------------------------------------------------------------------------
    # Component 1: Merged header in Sheet2 row 1 with blue fill and bold white text (0.25 pts)
    # This verifies the "Logistics Performance Report" merged header cell
    # A1:F1 merged, blue fill (FF4472C4), bold font, white font color
    # -------------------------------------------------------------------------
    try:
        header_cell = ws['A1']
        # Check header exists (non-empty text)
        has_header_text = (
            header_cell.value is not None
            and len(str(header_cell.value).strip()) > 0
        )

        # Check that A1:F1 is merged (look for A1:F1 or wider merge in merged_cells)
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if (merged_range.min_row == 1 and merged_range.max_row == 1
                    and merged_range.min_col == 1 and merged_range.max_col >= 2):
                is_merged = True
                break

        # Check blue fill (allow FF4472C4 or similar blue ARGB)
        has_blue_fill = False
        try:
            fgcolor = header_cell.fill.fgColor.rgb
            # Check for blue-spectrum fill: FF prefix with meaningful blue component
            # Expected: FF4472C4 but accept any fill that is not white/empty
            if fgcolor and fgcolor != '00000000' and fgcolor != 'FFFFFFFF' and fgcolor != '00000000':
                # Verify it is a solid fill (not 'none')
                if header_cell.fill.fill_type == 'solid':
                    has_blue_fill = True
        except Exception:
            has_blue_fill = False

        # Check bold font
        has_bold = header_cell.font.bold is True

        # Check white font color (00FFFFFF or FFFFFFFF)
        has_white_font = False
        try:
            font_rgb = header_cell.font.color.rgb
            # Accept 6-char or 8-char representations of white
            if font_rgb and (font_rgb.upper().endswith('FFFFFF') or font_rgb.upper() == '00FFFFFF'):
                has_white_font = True
        except Exception:
            has_white_font = False

        if has_header_text and is_merged and has_blue_fill and has_bold and has_white_font:
            print(f"PASS: Component 1 — Merged header present: '{header_cell.value}', "
                  f"blue fill ({header_cell.fill.fgColor.rgb}), bold={has_bold}, "
                  f"white font, merged range present (0.25 pts)")
            total_score += 0.25
        else:
            details = []
            if not has_header_text:
                details.append(f"no header text (A1={repr(header_cell.value)})")
            if not is_merged:
                details.append(f"A1 not merged across columns (ranges={list(ws.merged_cells.ranges)})")
            if not has_blue_fill:
                try:
                    details.append(f"no blue fill (fgColor={header_cell.fill.fgColor.rgb}, type={header_cell.fill.fill_type})")
                except Exception:
                    details.append("fill read error")
            if not has_bold:
                details.append("font not bold")
            if not has_white_font:
                try:
                    details.append(f"font not white (color={header_cell.font.color.rgb})")
                except Exception:
                    details.append("font color read error")
            print(f"FAIL: Component 1 — Merged styled header: {'; '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Pivot table 1 — Total Shipments by Carrier (0.25 pts)
    # Expects a section titled 'Total Shipments by Carrier' (or similar)
    # with carrier names (DHL, FedEx, UPS, USPS) and their shipment counts
    # Expected counts (computed from Sheet1 data):
    #   DHL=6, FedEx=7, UPS=6, USPS=6
    # -------------------------------------------------------------------------
    try:
        # Find pivot table 1 section by scanning Sheet2 for relevant header
        # and data cells. We allow flexible positioning — search all cells.
        pivot1_found = False
        pivot1_title_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if (cell.value is not None
                        and 'carrier' in str(cell.value).lower()
                        and 'shipment' in str(cell.value).lower()):
                    pivot1_title_row = cell.row
                    pivot1_found = True
                    break
            if pivot1_found:
                break

        if not pivot1_found:
            print("FAIL: Component 2 — No 'Total Shipments by Carrier' section found in Sheet2")
        else:
            # Scan rows below pivot1_title_row for carrier-count data
            # Allow up to 15 rows below title
            carrier_data = {}
            for scan_row in range(pivot1_title_row + 1, pivot1_title_row + 15):
                if scan_row > ws.max_row:
                    break
                cell_a = ws.cell(row=scan_row, column=1)
                cell_b = ws.cell(row=scan_row, column=2)
                if cell_a.value is not None and cell_b.value is not None:
                    try:
                        carrier_data[str(cell_a.value).strip()] = float(cell_b.value)
                    except (ValueError, TypeError):
                        pass

            # Validate expected carriers and counts
            expected_carriers = {
                'DHL': 6, 'FedEx': 7, 'UPS': 6, 'USPS': 6
            }
            carriers_correct = 0
            for carrier, expected_count in expected_carriers.items():
                if carrier in carrier_data:
                    actual = carrier_data[carrier]
                    if abs(actual - expected_count) < 0.5:  # integer counts
                        carriers_correct += 1

            if carriers_correct >= 3:  # at least 3 of 4 carriers correct
                print(f"PASS: Component 2 — Total Shipments by Carrier pivot found at row {pivot1_title_row}. "
                      f"Carrier data: {carrier_data} ({carriers_correct}/4 carriers correct) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — Carrier pivot found but data incorrect. "
                      f"Found: {carrier_data}, Expected: {expected_carriers} "
                      f"({carriers_correct}/4 correct)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Pivot table 2 — Average Delivery Time by Region (0.25 pts)
    # Expected averages (computed from Sheet1 data):
    #   East~4.57, North~4.0, South~5.83, West~4.2
    # -------------------------------------------------------------------------
    try:
        pivot2_found = False
        pivot2_title_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if (cell.value is not None
                        and 'region' in str(cell.value).lower()
                        and 'delivery' in str(cell.value).lower()):
                    pivot2_title_row = cell.row
                    pivot2_found = True
                    break
            if pivot2_found:
                break

        if not pivot2_found:
            print("FAIL: Component 3 — No 'Average Delivery Time by Region' section found in Sheet2")
        else:
            region_data = {}
            for scan_row in range(pivot2_title_row + 1, pivot2_title_row + 15):
                if scan_row > ws.max_row:
                    break
                cell_a = ws.cell(row=scan_row, column=1)
                cell_b = ws.cell(row=scan_row, column=2)
                if cell_a.value is not None and cell_b.value is not None:
                    try:
                        region_data[str(cell_a.value).strip()] = float(cell_b.value)
                    except (ValueError, TypeError):
                        pass

            # Validate expected regions and avg delivery times
            expected_regions = {
                'East': 4.57, 'North': 4.0, 'South': 5.83, 'West': 4.2
            }
            regions_correct = 0
            for region, expected_avg in expected_regions.items():
                if region in region_data:
                    actual = region_data[region]
                    if abs(actual - expected_avg) <= 0.1:  # tolerance for rounding
                        regions_correct += 1

            if regions_correct >= 3:  # at least 3 of 4 regions correct
                print(f"PASS: Component 3 — Avg Delivery Time by Region pivot found at row {pivot2_title_row}. "
                      f"Region data: {region_data} ({regions_correct}/4 regions correct) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Region pivot found but data incorrect. "
                      f"Found: {region_data}, Expected: {expected_regions} "
                      f"({regions_correct}/4 correct)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Pivot table 3 — Total Freight Cost by Shipment Type (0.25 pts)
    # Expected totals (computed from Sheet1 data):
    #   Express~2413.35, Freight~6200.65, Standard~1185.70
    # -------------------------------------------------------------------------
    try:
        pivot3_found = False
        pivot3_title_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if (cell.value is not None
                        and 'freight' in str(cell.value).lower()
                        and 'cost' in str(cell.value).lower()):
                    pivot3_title_row = cell.row
                    pivot3_found = True
                    break
            if pivot3_found:
                break

        if not pivot3_found:
            print("FAIL: Component 4 — No 'Total Freight Cost by Shipment Type' section found in Sheet2")
        else:
            freight_data = {}
            for scan_row in range(pivot3_title_row + 1, pivot3_title_row + 15):
                if scan_row > ws.max_row:
                    break
                cell_a = ws.cell(row=scan_row, column=1)
                cell_b = ws.cell(row=scan_row, column=2)
                if cell_a.value is not None and cell_b.value is not None:
                    try:
                        freight_data[str(cell_a.value).strip()] = float(cell_b.value)
                    except (ValueError, TypeError):
                        pass

            # Validate expected shipment types and totals
            expected_freight = {
                'Express': 2413.35, 'Freight': 6200.65, 'Standard': 1185.70
            }
            types_correct = 0
            for stype, expected_total in expected_freight.items():
                if stype in freight_data:
                    actual = freight_data[stype]
                    if abs(actual - expected_total) <= 1.0:  # allow small rounding difference
                        types_correct += 1

            if types_correct >= 2:  # at least 2 of 3 types correct
                print(f"PASS: Component 4 — Total Freight Cost by Shipment Type pivot found at row {pivot3_title_row}. "
                      f"Type data: {freight_data} ({types_correct}/3 types correct) (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 4 — Freight pivot found but data incorrect. "
                      f"Found: {freight_data}, Expected: {expected_freight} "
                      f"({types_correct}/3 correct)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint: save any open editor state, then run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'

persist_app_state()

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
