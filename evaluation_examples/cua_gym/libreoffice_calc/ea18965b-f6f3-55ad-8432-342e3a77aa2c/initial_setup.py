"""
Initial Setup: Income statement with net profit column empty, no Sheet2
Task ID: osworld_calc_gross_profit_sheet2_concat_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: IncomeStatement ---
    ws1 = wb.active
    ws1.title = 'IncomeStatement'

    # Headers: A=Fiscal Year, B=Revenue, C=COGS, D=Operating Expenses, E=Tax, F=Net Profit (empty)
    headers = ['Fiscal Year', 'Revenue', 'COGS', 'Operating Expenses', 'Tax', 'Net Profit']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Data rows — realistic business income statement data
    # (F column intentionally left empty — the task requires agent to fill it)
    data = [
        [2019, 4250000, 1870000, 820000, 185000],
        [2020, 3980000, 1750000, 795000, 162000],
        [2021, 4610000, 1990000, 855000, 198000],
        [2022, 5120000, 2230000, 910000, 224000],
        [2023, 5475000, 2380000, 965000, 241000],
        [2024, 5830000, 2510000, 1020000, 258000],
        [2025, 6190000, 2670000, 1085000, 278000],
        [2026, 6550000, 2820000, 1140000, 295000],
        [2027, 6920000, 2980000, 1205000, 316000],
        [2028, 7310000, 3145000, 1270000, 334000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
        # Column F (Net Profit) intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
