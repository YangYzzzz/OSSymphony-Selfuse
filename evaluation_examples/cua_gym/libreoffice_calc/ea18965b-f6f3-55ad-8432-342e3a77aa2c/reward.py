"""
Reward Script: Add net profit column formulas and Sheet2 summary text formula
Task ID: osworld_calc_gross_profit_sheet2_concat_005
Domain: libreoffice_calc
Scoring:
  Component 1: Net profit formulas in F2:F11 of IncomeStatement (0.5 pts)
  Component 2: Sheet2 exists in the workbook (0.2 pts)
  Component 3: Sheet2 A1 contains a TEXT concat formula for FY label + net profit (0.3 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_005'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Add a net profit column (F) with formulas =B#-C#-D#-E# for each data row,
    then create Sheet2 with A1 containing a formatted text concat formula like:
      ="FY"&<FY_cell>&"_Net Profit: $"&TEXT(<NetProfit_cell>,"#,##0.00")
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Net profit formulas in F column, rows 2-11 (0.5 points)
    # Each data row must have formula =B#-C#-D#-E# in column F
    try:
        ws_income = wb["IncomeStatement"] if "IncomeStatement" in wb.sheetnames else None
        if ws_income is None:
            print("FAIL: Component 1 — 'IncomeStatement' sheet not found")
        else:
            # Check F2:F11 for net profit formulas
            formula_pattern = re.compile(
                r'^=B\d+-C\d+-D\d+-E\d+$', re.IGNORECASE
            )
            data_rows = range(2, 12)  # rows 2 through 11
            formulas_found = 0
            formula_errors = []

            for row in data_rows:
                cell_val = ws_income.cell(row=row, column=6).value  # column F
                if cell_val is not None and isinstance(cell_val, str):
                    # Normalize: remove spaces
                    normalized = cell_val.replace(" ", "")
                    if formula_pattern.match(normalized):
                        # Also verify the row numbers match
                        expected = f"=B{row}-C{row}-D{row}-E{row}"
                        if normalized.upper() == expected.upper():
                            formulas_found += 1
                        else:
                            formula_errors.append(f"F{row}: formula row mismatch, got {cell_val}, expected {expected}")
                    else:
                        formula_errors.append(f"F{row}: formula pattern mismatch, got {repr(cell_val)}")
                else:
                    formula_errors.append(f"F{row}: empty or non-formula value, got {repr(cell_val)}")

            if formulas_found == 10:
                print(f"PASS: Component 1 — All 10 net profit formulas found in F2:F11 (0.5 pts)")
                total_score += 0.5
            elif formulas_found >= 5:
                # Partial credit for at least half correct
                print(f"PARTIAL: Component 1 — {formulas_found}/10 net profit formulas found (0.25 pts)")
                print(f"  Issues: {formula_errors[:3]}")
                if formulas_found >= 5:
                    total_score += 0.25
            else:
                print(f"FAIL: Component 1 — Only {formulas_found}/10 net profit formulas found (0 pts)")
                print(f"  Issues: {formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Sheet2 exists (0.2 points)
    try:
        if "Sheet2" in wb.sheetnames:
            print("PASS: Component 2 — Sheet2 exists in workbook (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Sheet2 not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet2 A1 contains TEXT concat formula for FY + net profit (0.3 points)
    # Formula must reference FY label and net profit with TEXT() formatting
    # Expected pattern: ="FY"&<ref>&"_Net Profit: $"&TEXT(<ref>,"#,##0.00")
    try:
        if "Sheet2" not in wb.sheetnames:
            print("FAIL: Component 3 — Sheet2 not found, cannot check A1 formula")
        else:
            ws_sheet2 = wb["Sheet2"]
            a1_val = ws_sheet2["A1"].value

            if a1_val is None:
                print("FAIL: Component 3 — Sheet2 A1 is empty")
            elif not isinstance(a1_val, str):
                print(f"FAIL: Component 3 — Sheet2 A1 is not a formula string, got: {repr(a1_val)}")
            else:
                formula_str = a1_val.upper().replace(" ", "")

                # Check key elements of the formula:
                # 1. Starts with = (is a formula)
                # 2. Contains "FY" string literal
                # 3. Uses TEXT() function
                # 4. Contains "#,##0.00" format string
                # 5. References Net Profit value (likely F2 or IncomeStatement.F2)
                # 6. Contains "_Net Profit: $" or similar text

                has_formula = formula_str.startswith("=")
                has_fy = '"FY"' in a1_val or "'FY'" in a1_val
                has_text_func = "TEXT(" in formula_str
                has_format = "#,##0.00" in a1_val
                has_net_profit_ref = (
                    "F2" in formula_str or
                    "INCOMESTATEMENT.F2" in formula_str or
                    "INCOMESTATEMENT!F2" in formula_str
                )
                # Check for "Net Profit" label in original formula (case-insensitive, with spaces preserved)
                has_net_profit_label = "net profit" in a1_val.lower() or "net_profit" in a1_val.lower()

                checks = {
                    "is formula (starts with =)": has_formula,
                    "contains FY string literal": has_fy,
                    "uses TEXT() function": has_text_func,
                    "uses #,##0.00 format": has_format,
                    "references Net Profit cell (F2)": has_net_profit_ref,
                    "contains Net Profit label": has_net_profit_label,
                }

                passed_checks = sum(1 for v in checks.values() if v)
                total_checks = len(checks)

                if passed_checks == total_checks:
                    print(f"PASS: Component 3 — Sheet2 A1 has complete TEXT concat formula (0.3 pts)")
                    print(f"  Formula: {a1_val}")
                    total_score += 0.3
                elif passed_checks >= 4:
                    print(f"PARTIAL: Component 3 — Sheet2 A1 formula partially correct ({passed_checks}/{total_checks} checks, 0.15 pts)")
                    print(f"  Formula: {a1_val}")
                    for check_name, passed in checks.items():
                        status = "PASS" if passed else "FAIL"
                        print(f"    {status}: {check_name}")
                    if passed_checks >= 4:
                        total_score += 0.15
                else:
                    print(f"FAIL: Component 3 — Sheet2 A1 formula missing key elements ({passed_checks}/{total_checks} checks)")
                    print(f"  Formula: {a1_val}")
                    for check_name, passed in checks.items():
                        status = "PASS" if passed else "FAIL"
                        print(f"    {status}: {check_name}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.1f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
