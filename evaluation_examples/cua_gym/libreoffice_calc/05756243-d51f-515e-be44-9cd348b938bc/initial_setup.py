"""
Initial Setup: Apply built-in cell styles (Heading 1 and Good) to specific cells
Task ID: calc_gg3_013
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Overview Sheet ---
    ws = wb.active
    ws.title = 'Overview'

    # Section 1: Revenue Summary
    ws.cell(row=1, column=1, value='Revenue Summary')
    ws.cell(row=1, column=2, value='Amount ($)')

    # Revenue data labels and values (B2:B8 range from 450 to 2800)
    revenue_data = [
        ['Online Sales', 2350],
        ['Retail Stores', 1850],
        ['Wholesale', 780],
        ['Subscription Services', 2800],
        ['Consulting Fees', 450],
        ['Licensing Revenue', 1200],
        ['Maintenance Contracts', 950],
    ]
    for r, (label, amount) in enumerate(revenue_data, 2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=amount)

    # Empty row 9 as separator
    ws.cell(row=9, column=1, value='')

    # Section 2: Cost Summary
    ws.cell(row=10, column=1, value='Cost Summary')
    ws.cell(row=10, column=2, value='Amount ($)')

    cost_data = [
        ['Salaries & Wages', 4200],
        ['Office Rent', 1500],
        ['Marketing Expenses', 890],
        ['Software Licenses', 620],
        ['Travel & Entertainment', 345],
        ['Utilities', 280],
        ['Insurance', 510],
    ]
    for r, (label, amount) in enumerate(cost_data, 11):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=amount)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
