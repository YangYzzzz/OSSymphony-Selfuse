"""
Reward Script: Apply 'Heading 1' cell style to A1 and A10, and 'Good' style to B2:B8 cells > 1000
Task ID: calc_gg3_013
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): A1 has Heading 1 style (bold, large font)
  Component 2 (0.25): A10 has Heading 1 style (bold, large font)
  Component 3 (0.30): Cells in B2:B8 with value > 1000 have 'Good' style (green fill FFCCFFCC)
  Component 4 (0.20): Cells in B2:B8 with value <= 1000 do NOT have the green fill (retain Default)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_013'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Overview' sheet must exist
    if 'Overview' not in wb.sheetnames:
        print("FAIL: 'Overview' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Overview']

    # Component 1: A1 has Heading 1 style applied (0.25 points)
    # Heading 1 in LibreOffice = bold, large font size (>=18), distinctive font
    # Initial state: Calibri 11, not bold -> should change to bold, large font
    try:
        cell_a1 = ws['A1']
        a1_bold = cell_a1.font.bold
        a1_size = cell_a1.font.size
        # Heading 1 style: bold=True and font size significantly larger than default (11)
        if a1_bold and a1_size is not None and a1_size >= 18:
            print(f"PASS: Component 1 — A1 has Heading 1 style (bold={a1_bold}, size={a1_size}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — A1 expected bold=True and size>=18, found bold={a1_bold}, size={a1_size}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: A10 has Heading 1 style applied (0.25 points)
    try:
        cell_a10 = ws['A10']
        a10_bold = cell_a10.font.bold
        a10_size = cell_a10.font.size
        if a10_bold and a10_size is not None and a10_size >= 18:
            print(f"PASS: Component 2 — A10 has Heading 1 style (bold={a10_bold}, size={a10_size}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — A10 expected bold=True and size>=18, found bold={a10_bold}, size={a10_size}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Cells in B2:B8 with value > 1000 have 'Good' style (green fill) (0.30 points)
    # Good style in LibreOffice = solid green fill (FFCCFFCC or similar green)
    # Values > 1000: B2(2350), B3(1850), B5(2800), B7(1200) = 4 cells
    try:
        cells_gt_1000 = []
        for row in range(2, 9):
            cell = ws.cell(row=row, column=2)
            val = cell.value
            if val is not None and isinstance(val, (int, float)) and val > 1000:
                cells_gt_1000.append((row, cell))

        if len(cells_gt_1000) == 0:
            print("FAIL: Component 3 — No cells with value > 1000 found in B2:B8")
        else:
            good_count = 0
            for row, cell in cells_gt_1000:
                fill_type = cell.fill.fill_type
                fill_rgb = None
                try:
                    fill_rgb = cell.fill.fgColor.rgb
                except:
                    pass
                # Check for green-ish fill (Good style uses CCFFCC green or similar)
                has_green_fill = (
                    fill_type == 'solid' and
                    fill_rgb is not None and
                    fill_rgb != '00000000' and
                    # Accept various green fill colors used by 'Good' style
                    ('CCFF' in fill_rgb or 'ccff' in fill_rgb.lower() or
                     # Also accept if green channel is dominant
                     _is_green_fill(fill_rgb))
                )
                if has_green_fill:
                    good_count += 1
                    print(f"  B{row} (val={cell.value}): Good style applied (fill={fill_rgb})")
                else:
                    print(f"  B{row} (val={cell.value}): Missing Good style (fill_type={fill_type}, fill_rgb={fill_rgb})")

            ratio = good_count / len(cells_gt_1000)
            points = round(0.30 * ratio, 4)
            if good_count == len(cells_gt_1000):
                print(f"PASS: Component 3 — All {good_count} cells > 1000 have Good style ({points} pts)")
            else:
                print(f"PARTIAL: Component 3 — {good_count}/{len(cells_gt_1000)} cells > 1000 have Good style ({points} pts)")
            total_score += points
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Cells in B2:B8 with value <= 1000 do NOT have green fill (0.20 points)
    # This ensures selectivity: only cells > 1000 were styled, not all cells.
    # IMPORTANT: Only award these points if at least one cell > 1000 actually has Good style,
    # otherwise this check is a precondition (true before task) not a task-change check.
    # Values <= 1000: B4(780), B6(450), B8(950) = 3 cells
    try:
        # Gate: only score selectivity if some Good style was actually applied (Component 3 scored > 0)
        any_good_applied = False
        for row in range(2, 9):
            cell = ws.cell(row=row, column=2)
            val = cell.value
            if val is not None and isinstance(val, (int, float)) and val > 1000:
                ft = cell.fill.fill_type
                fr = None
                try:
                    fr = cell.fill.fgColor.rgb
                except:
                    pass
                if ft == 'solid' and fr is not None and fr != '00000000':
                    any_good_applied = True
                    break

        if not any_good_applied:
            print("SKIP: Component 4 — No Good style applied to any cell, selectivity check not applicable (0 pts)")
        else:
            cells_lte_1000 = []
            for row in range(2, 9):
                cell = ws.cell(row=row, column=2)
                val = cell.value
                if val is not None and isinstance(val, (int, float)) and val <= 1000:
                    cells_lte_1000.append((row, cell))

            if len(cells_lte_1000) == 0:
                print("FAIL: Component 4 — No cells with value <= 1000 found in B2:B8")
            else:
                correct_count = 0
                for row, cell in cells_lte_1000:
                    fill_type = cell.fill.fill_type
                    fill_rgb = None
                    try:
                        fill_rgb = cell.fill.fgColor.rgb
                    except:
                        pass
                    # These should NOT have a green solid fill
                    has_no_green = (
                        fill_type is None or
                        fill_type != 'solid' or
                        fill_rgb is None or
                        fill_rgb == '00000000'
                    )
                    if has_no_green:
                        correct_count += 1
                        print(f"  B{row} (val={cell.value}): Correctly no Good style")
                    else:
                        print(f"  B{row} (val={cell.value}): Incorrectly has fill (fill_type={fill_type}, fill_rgb={fill_rgb})")

                ratio = correct_count / len(cells_lte_1000)
                points = round(0.20 * ratio, 4)
                if correct_count == len(cells_lte_1000):
                    print(f"PASS: Component 4 — All {correct_count} cells <= 1000 retain Default style ({points} pts)")
                else:
                    print(f"PARTIAL: Component 4 — {correct_count}/{len(cells_lte_1000)} cells <= 1000 correctly unstyled ({points} pts)")
                total_score += points
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


def _is_green_fill(rgb_str):
    """Check if an ARGB hex string represents a green-ish fill color."""
    if not rgb_str or len(rgb_str) < 6:
        return False
    # Extract RGB components (last 6 chars for ARGB, or all 6 for RGB)
    hex_part = rgb_str[-6:]
    try:
        r = int(hex_part[0:2], 16)
        g = int(hex_part[2:4], 16)
        b = int(hex_part[4:6], 16)
        # Green fill: green channel should be high, and higher than red and blue
        return g > 150 and g > r and g > b
    except ValueError:
        return False


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
