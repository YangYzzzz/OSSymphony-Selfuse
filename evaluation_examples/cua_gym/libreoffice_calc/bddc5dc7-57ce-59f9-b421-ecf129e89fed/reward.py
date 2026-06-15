"""
Reward Script: Create two pivot tables (ByRegion, ByProduct) from SourceData
Task ID: calc_pivot_053
Domain: libreoffice_calc
Scoring:
  - Component 1: ByRegion sheet exists (0.15)
  - Component 2: ByProduct sheet exists (0.15)
  - Component 3: ByRegion headers correct (0.10)
  - Component 4: ByProduct headers correct (0.10)
  - Component 5: ByRegion revenue values correct (0.20)
  - Component 6: ByProduct revenue values correct (0.20)
  - Component 7: Both grand totals equal 280000 (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_053'


def persist_app_state(domain: str):
    """Attempt to save any unsaved LibreOffice edits."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_value_approx(actual, expected, tolerance=0.01):
    """Check if numeric value matches within tolerance."""
    if actual is None:
        return False
    try:
        return abs(float(actual) - float(expected)) <= tolerance
    except (ValueError, TypeError):
        return False


def find_value_in_column(ws, col_label, col_value, target_col, max_row=20):
    """Find a row where col_label column has col_value, return target_col value.
    Searches by scanning the sheet for label-value pairs."""
    for row in range(2, max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        if cell_val is not None and str(cell_val).strip().lower() == str(col_value).strip().lower():
            return ws.cell(row=row, column=2).value
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Found sheets: {sheet_names}")

    # Component 1: ByRegion sheet exists (0.15 points)
    try:
        if 'ByRegion' in sheet_names:
            print(f"PASS: Component 1 - ByRegion sheet exists (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 - ByRegion sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: ByProduct sheet exists (0.15 points)
    try:
        if 'ByProduct' in sheet_names:
            print(f"PASS: Component 2 - ByProduct sheet exists (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 - ByProduct sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: ByRegion headers correct (0.10 points)
    try:
        if 'ByRegion' in sheet_names:
            ws_region = wb['ByRegion']
            header_a = ws_region.cell(row=1, column=1).value
            header_b = ws_region.cell(row=1, column=2).value
            # Accept flexible header naming
            region_header_ok = header_a is not None and 'region' in str(header_a).strip().lower()
            revenue_header_ok = header_b is not None and 'revenue' in str(header_b).strip().lower()
            if region_header_ok and revenue_header_ok:
                print(f"PASS: Component 3 - ByRegion headers: '{header_a}', '{header_b}' (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 - ByRegion headers: '{header_a}', '{header_b}' (expected Region/Revenue)")
        else:
            print(f"FAIL: Component 3 - ByRegion sheet missing, cannot check headers")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: ByProduct headers correct (0.10 points)
    try:
        if 'ByProduct' in sheet_names:
            ws_product = wb['ByProduct']
            header_a = ws_product.cell(row=1, column=1).value
            header_b = ws_product.cell(row=1, column=2).value
            product_header_ok = header_a is not None and 'product' in str(header_a).strip().lower()
            revenue_header_ok = header_b is not None and 'revenue' in str(header_b).strip().lower()
            if product_header_ok and revenue_header_ok:
                print(f"PASS: Component 4 - ByProduct headers: '{header_a}', '{header_b}' (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 - ByProduct headers: '{header_a}', '{header_b}' (expected Product/Revenue)")
        else:
            print(f"FAIL: Component 4 - ByProduct sheet missing, cannot check headers")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: ByRegion revenue values correct (0.20 points)
    # Expected: North=82000, South=68000, East=75000, West=55000
    try:
        if 'ByRegion' in sheet_names:
            ws_region = wb['ByRegion']
            expected_region = {
                'north': 82000,
                'south': 68000,
                'east': 75000,
                'west': 55000,
            }
            region_matches = 0
            for region_name, expected_val in expected_region.items():
                actual_val = find_value_in_column(ws_region, 'Region', region_name, 'Revenue')
                if actual_val is not None and check_value_approx(actual_val, expected_val, tolerance=1.0):
                    region_matches += 1
                    print(f"  INFO: ByRegion {region_name}={actual_val} (expected {expected_val}) OK")
                else:
                    print(f"  INFO: ByRegion {region_name}={actual_val} (expected {expected_val}) MISMATCH")

            if region_matches == 4:
                print(f"PASS: Component 5 - All 4 region revenue values correct (0.20 pts)")
                total_score += 0.20
            elif region_matches >= 2:
                partial = round(0.20 * region_matches / 4, 2)
                print(f"PARTIAL: Component 5 - {region_matches}/4 region values correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 - Only {region_matches}/4 region values correct")
        else:
            print(f"FAIL: Component 5 - ByRegion sheet missing")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: ByProduct revenue values correct (0.20 points)
    # Expected: X=110000, Y=95000, Z=75000
    try:
        if 'ByProduct' in sheet_names:
            ws_product = wb['ByProduct']
            expected_product = {
                'x': 110000,
                'y': 95000,
                'z': 75000,
            }
            product_matches = 0
            for product_name, expected_val in expected_product.items():
                actual_val = find_value_in_column(ws_product, 'Product', product_name, 'Revenue')
                if actual_val is not None and check_value_approx(actual_val, expected_val, tolerance=1.0):
                    product_matches += 1
                    print(f"  INFO: ByProduct {product_name}={actual_val} (expected {expected_val}) OK")
                else:
                    print(f"  INFO: ByProduct {product_name}={actual_val} (expected {expected_val}) MISMATCH")

            if product_matches == 3:
                print(f"PASS: Component 6 - All 3 product revenue values correct (0.20 pts)")
                total_score += 0.20
            elif product_matches >= 1:
                partial = round(0.20 * product_matches / 3, 2)
                print(f"PARTIAL: Component 6 - {product_matches}/3 product values correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 6 - Only {product_matches}/3 product values correct")
        else:
            print(f"FAIL: Component 6 - ByProduct sheet missing")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Both grand totals equal 280000 (0.10 points)
    try:
        grand_total_ok = 0
        if 'ByRegion' in sheet_names:
            ws_region = wb['ByRegion']
            gt_val = find_value_in_column(ws_region, 'label', 'grand total', 'value')
            if gt_val is not None and check_value_approx(gt_val, 280000, tolerance=1.0):
                grand_total_ok += 1
                print(f"  INFO: ByRegion Grand Total={gt_val} OK")
            else:
                # Also check the last data row
                for r in range(ws_region.max_row, 1, -1):
                    v = ws_region.cell(row=r, column=2).value
                    label = ws_region.cell(row=r, column=1).value
                    if v is not None and check_value_approx(v, 280000, tolerance=1.0):
                        grand_total_ok += 1
                        print(f"  INFO: ByRegion Grand Total at row {r}={v} OK")
                        break
                else:
                    print(f"  INFO: ByRegion Grand Total not found or incorrect")

        if 'ByProduct' in sheet_names:
            ws_product = wb['ByProduct']
            gt_val = find_value_in_column(ws_product, 'label', 'grand total', 'value')
            if gt_val is not None and check_value_approx(gt_val, 280000, tolerance=1.0):
                grand_total_ok += 1
                print(f"  INFO: ByProduct Grand Total={gt_val} OK")
            else:
                for r in range(ws_product.max_row, 1, -1):
                    v = ws_product.cell(row=r, column=2).value
                    label = ws_product.cell(row=r, column=1).value
                    if v is not None and check_value_approx(v, 280000, tolerance=1.0):
                        grand_total_ok += 1
                        print(f"  INFO: ByProduct Grand Total at row {r}={v} OK")
                        break
                else:
                    print(f"  INFO: ByProduct Grand Total not found or incorrect")

        if grand_total_ok == 2:
            print(f"PASS: Component 7 - Both grand totals equal 280000 (0.10 pts)")
            total_score += 0.10
        elif grand_total_ok == 1:
            print(f"PARTIAL: Component 7 - Only 1/2 grand totals correct (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 - No grand totals found matching 280000")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
