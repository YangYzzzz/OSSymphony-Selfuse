"""
Initial Setup: Create source data for pivot table task
Task ID: calc_pivot_053
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SourceData'

    # Headers
    headers = ['ID', 'Region', 'Product', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Define the target sums per (region, product) combo
    combos = {
        ('North', 'X'): 32213,
        ('North', 'Y'): 27822,
        ('North', 'Z'): 21965,
        ('South', 'X'): 26715,
        ('South', 'Y'): 23071,
        ('South', 'Z'): 18214,
        ('East', 'X'): 29465,
        ('East', 'Y'): 25446,
        ('East', 'Z'): 20089,
        ('West', 'X'): 21607,
        ('West', 'Y'): 18661,
        ('West', 'Z'): 14732,
    }

    # Distribute 200 rows across 12 combos
    # Roughly equal: 16-17 rows per combo
    combo_keys = list(combos.keys())
    rows_per_combo = {}
    base = 200 // 12  # 16
    remainder = 200 % 12  # 8
    for i, key in enumerate(combo_keys):
        rows_per_combo[key] = base + (1 if i < remainder else 0)

    # Generate rows
    all_rows = []
    for (region, product), total_revenue in combos.items():
        n_rows = rows_per_combo[(region, product)]
        # Split total_revenue into n_rows random positive integers
        if n_rows == 1:
            values = [total_revenue]
        else:
            # Generate n_rows-1 random breakpoints
            base_val = total_revenue // n_rows
            values = [base_val] * n_rows
            leftover = total_revenue - sum(values)
            # Distribute leftover randomly
            for _ in range(leftover):
                idx = random.randint(0, n_rows - 1)
                values[idx] += 1
            # Add some variance
            for i in range(n_rows - 1):
                shift = random.randint(-base_val // 3, base_val // 3)
                values[i] += shift
                values[i + 1] -= shift
                # Ensure positive
                if values[i] < 100:
                    values[i + 1] -= (100 - values[i])
                    values[i] = 100
                if values[i + 1] < 100:
                    values[i] -= (100 - values[i + 1])
                    values[i + 1] = 100

            # Final adjustment to ensure exact sum
            diff = total_revenue - sum(values)
            values[-1] += diff

        for v in values:
            all_rows.append((region, product, v))

    # Shuffle for realism
    random.shuffle(all_rows)

    # Write rows
    for i, (region, product, revenue) in enumerate(all_rows, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=region)
        ws.cell(row=row, column=3, value=product)
        ws.cell(row=row, column=4, value=revenue)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify sums
    region_sums = {}
    product_sums = {}
    total = 0
    for region, product, revenue in all_rows:
        region_sums[region] = region_sums.get(region, 0) + revenue
        product_sums[product] = product_sums.get(product, 0) + revenue
        total += revenue
    print(f'Region sums: {region_sums}')
    print(f'Product sums: {product_sums}')
    print(f'Grand total: {total}')
    print(f'Total rows: {len(all_rows)}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
