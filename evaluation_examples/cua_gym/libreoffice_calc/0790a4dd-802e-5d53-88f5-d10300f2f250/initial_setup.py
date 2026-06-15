"""
Initial Setup: Configure print settings for Summary sheet
Task ID: calc_gg3_039
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary (main sheet with A1:H30) ---
    ws = wb.active
    ws.title = 'Summary'

    headers = [
        'Employee Name', 'Department', 'Q1 Revenue', 'Q2 Revenue',
        'Q3 Revenue', 'Q4 Revenue', 'Annual Total', 'Rating'
    ]

    # Header styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows (29 rows of realistic employee data)
    data = [
        ['Sarah Chen', 'Engineering', 45230, 51200, 48900, 52100, 197430, 'Exceeds'],
        ['Marcus Johnson', 'Marketing', 38400, 42100, 39800, 44500, 164800, 'Meets'],
        ['Priya Patel', 'Sales', 62300, 58700, 67200, 71400, 259600, 'Exceeds'],
        ['David Kim', 'Engineering', 41500, 43800, 45200, 46900, 177400, 'Meets'],
        ['Rachel Torres', 'Finance', 35600, 37200, 36800, 38100, 147700, 'Meets'],
        ['James O\'Brien', 'Sales', 55800, 61200, 58400, 63700, 239100, 'Exceeds'],
        ['Aisha Mohammed', 'HR', 28900, 30100, 29500, 31200, 119700, 'Meets'],
        ['Carlos Rivera', 'Operations', 33700, 35400, 34200, 36800, 140100, 'Meets'],
        ['Emily Watson', 'Engineering', 47800, 49200, 51300, 53600, 201900, 'Exceeds'],
        ['Hiroshi Tanaka', 'Research', 42100, 44300, 43800, 45900, 176100, 'Meets'],
        ['Maria Gonzalez', 'Marketing', 36200, 38500, 37100, 40200, 152000, 'Meets'],
        ['Robert Singh', 'Sales', 58900, 62400, 60100, 65800, 247200, 'Exceeds'],
        ['Lisa Chang', 'Finance', 34100, 35800, 36200, 37500, 143600, 'Meets'],
        ['Thomas Mueller', 'Operations', 31200, 33400, 32100, 34700, 131400, 'Below'],
        ['Fatima Al-Hassan', 'Engineering', 44600, 46800, 48100, 50200, 189700, 'Meets'],
        ['Kevin Park', 'Research', 39800, 41200, 40500, 43100, 164600, 'Meets'],
        ['Sofia Petrov', 'HR', 27500, 29200, 28800, 30500, 116000, 'Meets'],
        ['Daniel Foster', 'Sales', 51200, 54800, 53100, 57400, 216500, 'Exceeds'],
        ['Yuki Nakamura', 'Marketing', 34800, 36500, 35900, 38200, 145400, 'Meets'],
        ['Amanda Brooks', 'Finance', 36900, 38400, 37800, 39600, 152700, 'Meets'],
        ['Omar Hassan', 'Engineering', 43200, 45100, 46800, 48300, 183400, 'Meets'],
        ['Isabella Rossi', 'Operations', 30800, 32600, 31900, 33500, 128800, 'Below'],
        ['Michael Lee', 'Research', 40500, 42800, 41900, 44200, 169400, 'Meets'],
        ['Natasha Volkov', 'Sales', 57100, 60300, 58800, 64100, 240300, 'Exceeds'],
        ['Benjamin Scott', 'HR', 29400, 31000, 30200, 32100, 122700, 'Meets'],
        ['Alicia Fernandez', 'Marketing', 37100, 39200, 38500, 41300, 156100, 'Meets'],
        ['Ryan O\'Connor', 'Finance', 35200, 37000, 36400, 38800, 147400, 'Meets'],
        ['Mei-Lin Wu', 'Engineering', 46100, 48300, 49700, 51800, 195900, 'Exceeds'],
        ['Patrick Dubois', 'Operations', 32400, 34100, 33500, 35900, 135900, 'Meets'],
    ]

    # Number formatting
    currency_fmt = '$#,##0'
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Name
        ws.cell(row=r, column=2, value=row_data[1])  # Department
        for c in range(3, 8):  # Q1-Q4 and Annual Total (columns C-G)
            cell = ws.cell(row=r, column=c, value=row_data[c - 1])
            cell.number_format = currency_fmt
        ws.cell(row=r, column=8, value=row_data[7])  # Rating

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions['H'].width = 12

    # --- Sheet 2: Departments ---
    ws2 = wb.create_sheet('Departments')
    dept_headers = ['Department', 'Head Count', 'Budget', 'Location']
    for col, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    dept_data = [
        ['Engineering', 6, 1200000, 'Building A'],
        ['Marketing', 3, 450000, 'Building B'],
        ['Sales', 4, 600000, 'Building C'],
        ['Finance', 3, 380000, 'Building A'],
        ['HR', 3, 290000, 'Building B'],
        ['Operations', 3, 350000, 'Building D'],
        ['Research', 3, 420000, 'Building A'],
    ]
    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Sheet 3: Notes ---
    ws3 = wb.create_sheet('Notes')
    ws3['A1'] = 'Annual Review Notes'
    ws3['A1'].font = Font(size=14, bold=True)
    ws3['A3'] = 'Review Period: January 2025 - December 2025'
    ws3['A4'] = 'Prepared by: HR Department'
    ws3['A5'] = 'Date: March 2026'
    ws3['A7'] = 'Rating Scale:'
    ws3['A8'] = 'Exceeds - Consistently exceeds performance expectations'
    ws3['A9'] = 'Meets - Meets all performance expectations'
    ws3['A10'] = 'Below - Performance improvement needed'

    # Ensure default print settings (no centering, no headings, no gridlines, no footer)
    # openpyxl defaults already have these off, but be explicit
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False
    ws.print_options.headings = False
    ws.print_options.gridLines = False
    # No header/footer configured by default

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
