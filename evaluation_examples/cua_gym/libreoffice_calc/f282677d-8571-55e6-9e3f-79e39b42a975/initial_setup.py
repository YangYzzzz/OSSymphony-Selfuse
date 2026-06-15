"""
Initial Setup: Create commission rate table and sales rep data for VLOOKUP task
Task ID: calc_sales_045
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 'Rates': Commission rate table ---
    ws_rates = wb.active
    ws_rates.title = 'Rates'

    # Headers
    headers_rates = ['Tier', 'Min Revenue', 'Rate']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers_rates, 1):
        cell = ws_rates.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # Rate data
    rate_data = [
        ['Bronze', 0, 0.05],
        ['Silver', 100000, 0.07],
        ['Gold', 250000, 0.09],
        ['Platinum', 500000, 0.12],
    ]
    for r, row_data in enumerate(rate_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_rates.cell(row=r, column=c, value=val)
            if c == 2:
                cell.number_format = '$#,##0'
            elif c == 3:
                cell.number_format = '0%'

    # Column widths for Rates
    ws_rates.column_dimensions['A'].width = 14
    ws_rates.column_dimensions['B'].width = 16
    ws_rates.column_dimensions['C'].width = 12

    # --- Sheet 'Reps': Sales rep data ---
    ws_reps = wb.create_sheet('Reps')

    headers_reps = ['Rep', 'Revenue', 'Rate', 'Commission']
    for col, h in enumerate(headers_reps, 1):
        cell = ws_reps.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Rep data - only names and revenue; Rate and Commission columns left empty
    rep_data = [
        ['Alice', 175000],
        ['Bob', 320000],
        ['Carol', 85000],
        ['Dan', 550000],
    ]
    for r, row_data in enumerate(rep_data, 2):
        ws_reps.cell(row=r, column=1, value=row_data[0])
        cell_rev = ws_reps.cell(row=r, column=2, value=row_data[1])
        cell_rev.number_format = '$#,##0'
        # C and D columns intentionally left empty (task is to fill them)

    # Column widths for Reps
    ws_reps.column_dimensions['A'].width = 14
    ws_reps.column_dimensions['B'].width = 16
    ws_reps.column_dimensions['C'].width = 12
    ws_reps.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
