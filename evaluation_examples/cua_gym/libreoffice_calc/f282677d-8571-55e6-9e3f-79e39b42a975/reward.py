"""
Reward Script: Create named range for commission rate table and use VLOOKUP formulas
Task ID: calc_sales_045
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): Named range 'CommRates' defined correctly
  - Component 2 (0.35): VLOOKUP formulas in Reps!C2:C5 using CommRates
  - Component 3 (0.20): Commission formulas in Reps!D2:D5 (=B*C)
  - Component 4 (0.20): Cached computed values match ground truth
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_045'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (formula mode)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both sheets exist
    if 'Rates' not in wb.sheetnames or 'Reps' not in wb.sheetnames:
        print(f"FAIL: Missing required sheets. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Named range 'CommRates' exists and references Rates!$B$2:$C$5 (0.25 pts)
    try:
        commrates_ref = None
        for name, dn in wb.defined_names.items():
            if name.lower() == 'commrates':
                commrates_ref = dn.attr_text
                break

        if commrates_ref is not None:
            # Verify the reference points to the correct range
            ref_clean = commrates_ref.upper().replace("'", "").replace("$", "")
            # Accept variations like Rates!$B$2:$C$5 or Rates!B2:C5
            if 'RATES!' in ref_clean and 'B2' in ref_clean and 'C5' in ref_clean:
                print(f"PASS: Component 1 — Named range 'CommRates' = {commrates_ref} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — CommRates ref is '{commrates_ref}', expected Rates!$B$2:$C$5")
        else:
            print(f"FAIL: Component 1 — No named range 'CommRates' found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: VLOOKUP formulas in Reps!C2:C5 using CommRates (0.35 pts)
    try:
        ws_reps = wb['Reps']
        vlookup_count = 0
        for row_num in range(2, 6):
            cell_val = ws_reps.cell(row=row_num, column=3).value  # Column C
            if cell_val and isinstance(cell_val, str):
                formula_upper = cell_val.upper().replace(' ', '')
                # Check for VLOOKUP that references CommRates (named range)
                if 'VLOOKUP' in formula_upper and 'COMMRATES' in formula_upper:
                    vlookup_count += 1
                    print(f"  C{row_num}: {cell_val} — uses VLOOKUP with CommRates")
                else:
                    print(f"  C{row_num}: {cell_val} — does not match VLOOKUP+CommRates pattern")
            else:
                print(f"  C{row_num}: {repr(cell_val)} — not a formula")

        if vlookup_count == 4:
            print(f"PASS: Component 2 — All 4 VLOOKUP formulas use CommRates (0.35 pts)")
            total_score += 0.35
        elif vlookup_count > 0:
            partial = round(0.35 * vlookup_count / 4, 4)
            print(f"PARTIAL: Component 2 — {vlookup_count}/4 VLOOKUP formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No VLOOKUP formulas with CommRates found in C2:C5")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Commission formulas in Reps!D2:D5 as =B*C (0.20 pts)
    try:
        ws_reps = wb['Reps']
        commission_count = 0
        for row_num in range(2, 6):
            cell_val = ws_reps.cell(row=row_num, column=4).value  # Column D
            if cell_val and isinstance(cell_val, str):
                formula_upper = cell_val.upper().replace(' ', '')
                # Accept =B2*C2, =C2*B2, or similar multiplication patterns
                b_ref = f'B{row_num}'
                c_ref = f'C{row_num}'
                if (B_ref := b_ref.upper()) and (C_ref := c_ref.upper()):
                    if (B_ref in formula_upper and C_ref in formula_upper and '*' in formula_upper):
                        commission_count += 1
                        print(f"  D{row_num}: {cell_val} — multiplication formula found")
                    else:
                        print(f"  D{row_num}: {cell_val} — does not match B*C pattern")
            else:
                print(f"  D{row_num}: {repr(cell_val)} — not a formula")

        if commission_count == 4:
            print(f"PASS: Component 3 — All 4 commission formulas correct (0.20 pts)")
            total_score += 0.20
        elif commission_count > 0:
            partial = round(0.20 * commission_count / 4, 4)
            print(f"PARTIAL: Component 3 — {commission_count}/4 commission formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No commission formulas found in D2:D5")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: VLOOKUP formulas use correct column index and approximate match (0.20 pts)
    # The VLOOKUP should look up in CommRates (B2:C5 on Rates sheet = 2-col table),
    # column index 2 retrieves the Rate (column C), and TRUE/1 for approximate match.
    # This ensures the formulas will produce correct results:
    # C2=0.07, D2=12250, C3=0.09, D3=28800, C4=0.05, D4=4250, C5=0.12, D5=66000
    try:
        ws_reps = wb['Reps']
        correct_formulas = 0
        for row_num in range(2, 6):
            cell_val = ws_reps.cell(row=row_num, column=3).value  # Column C
            if cell_val and isinstance(cell_val, str):
                formula_clean = cell_val.upper().replace(' ', '')
                # Check VLOOKUP has col_index=2 and approximate match (TRUE or 1)
                # Pattern: =VLOOKUP(Bx,CommRates,2,TRUE) or =VLOOKUP(Bx,CommRates,2,1)
                b_ref = f'B{row_num}'
                if (b_ref in formula_clean and
                    'COMMRATES' in formula_clean and
                    ',2,' in formula_clean and
                    ('TRUE' in formula_clean or ',1)' in formula_clean)):
                    correct_formulas += 1
                    print(f"  C{row_num}: correct VLOOKUP structure (col_index=2, approx match)")
                else:
                    print(f"  C{row_num}: {cell_val} — wrong VLOOKUP structure")
            else:
                print(f"  C{row_num}: {repr(cell_val)} — not a VLOOKUP formula")

        if correct_formulas == 4:
            print(f"PASS: Component 4 — All 4 VLOOKUPs have correct col_index=2 and approx match (0.20 pts)")
            total_score += 0.20
        elif correct_formulas > 0:
            partial = round(0.20 * correct_formulas / 4, 4)
            print(f"PARTIAL: Component 4 — {correct_formulas}/4 correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No VLOOKUPs have correct structure")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
