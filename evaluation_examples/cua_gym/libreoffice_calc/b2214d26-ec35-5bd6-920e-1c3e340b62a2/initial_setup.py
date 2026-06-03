"""
Initial Setup: Configure mailing label print settings
Task ID: calc_mcp_090
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Labels ---
    ws = wb.active
    ws.title = 'Labels'

    # Define label data: 3 columns of labels across, multiple rows of labels down
    # Each "label" block is 5 rows: Name, Address, City/State/Zip, blank, blank
    # Columns: A-C for label 1, E-G for label 2, I-K for label 3

    label_data = [
        {
            'name': 'Sarah Chen',
            'company': 'TechVision Labs',
            'address': '1425 Innovation Drive, Suite 300',
            'city_state_zip': 'San Francisco, CA 94107',
        },
        {
            'name': 'Marcus Johnson',
            'company': 'Greenfield Consulting',
            'address': '782 Oak Boulevard',
            'city_state_zip': 'Portland, OR 97201',
        },
        {
            'name': 'Priya Patel',
            'company': 'Horizon Medical Group',
            'address': '3500 Wellness Parkway, Bldg C',
            'city_state_zip': 'Austin, TX 78701',
        },
        {
            'name': 'David Kim',
            'company': 'Apex Financial Services',
            'address': '900 Commerce Street, Floor 12',
            'city_state_zip': 'Chicago, IL 60601',
        },
        {
            'name': 'Elena Rodriguez',
            'company': 'Starlight Education Foundation',
            'address': '215 Learning Lane',
            'city_state_zip': 'Denver, CO 80202',
        },
        {
            'name': 'James O\'Brien',
            'company': 'Pacific Rim Imports',
            'address': '4801 Harbor View Road',
            'city_state_zip': 'Seattle, WA 98101',
        },
        {
            'name': 'Aisha Washington',
            'company': 'Bright Futures Nonprofit',
            'address': '1122 Community Circle',
            'city_state_zip': 'Atlanta, GA 30301',
        },
        {
            'name': 'Robert Nakamura',
            'company': 'Summit Architecture',
            'address': '667 Design Plaza, Suite 8',
            'city_state_zip': 'Boston, MA 02101',
        },
        {
            'name': 'Lisa Fernandez',
            'company': 'Coastal Realty Group',
            'address': '3290 Beachside Avenue',
            'city_state_zip': 'Miami, FL 33101',
        },
        {
            'name': 'Thomas Wright',
            'company': 'Midwest Manufacturing Co.',
            'address': '508 Industrial Parkway',
            'city_state_zip': 'Detroit, MI 48201',
        },
        {
            'name': 'Jennifer Chang',
            'company': 'Digital Canvas Media',
            'address': '1750 Creative Way, Unit 4B',
            'city_state_zip': 'Los Angeles, CA 90001',
        },
        {
            'name': 'Michael Torres',
            'company': 'Evergreen Landscaping',
            'address': '432 Garden Path Drive',
            'city_state_zip': 'Phoenix, AZ 85001',
        },
        {
            'name': 'Amanda Foster',
            'company': 'BlueSky Analytics',
            'address': '2100 Data Drive, Floor 5',
            'city_state_zip': 'Raleigh, NC 27601',
        },
        {
            'name': 'Ryan Sullivan',
            'company': 'Heritage Construction LLC',
            'address': '855 Builder Lane',
            'city_state_zip': 'Nashville, TN 37201',
        },
        {
            'name': 'Sophia Martinez',
            'company': 'Luna Pet Care',
            'address': '1340 Paw Print Avenue',
            'city_state_zip': 'San Antonio, TX 78201',
        },
    ]

    # Layout: 3 labels per row, separated by blank columns
    # Label columns: A(1), D(4), G(7) — with B,C / E,F / H,I as extra space
    # Each label block = 5 rows (name, company, address, city/zip, blank separator)

    label_font_name = Font(name='Arial', size=10, bold=True)
    label_font_normal = Font(name='Arial', size=9)

    col_starts = [1, 4, 7]  # columns A, D, G

    row = 1
    label_idx = 0
    while label_idx < len(label_data):
        for col_offset, col_start in enumerate(col_starts):
            if label_idx >= len(label_data):
                break
            entry = label_data[label_idx]
            ws.cell(row=row, column=col_start, value=entry['name']).font = label_font_name
            ws.cell(row=row + 1, column=col_start, value=entry['company']).font = label_font_normal
            ws.cell(row=row + 2, column=col_start, value=entry['address']).font = label_font_normal
            ws.cell(row=row + 3, column=col_start, value=entry['city_state_zip']).font = label_font_normal
            label_idx += 1
        row += 5  # 4 data rows + 1 blank separator row

    # Set column widths for label layout
    for col_start in col_starts:
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_start)].width = 32

    # Set a print area (this is what the task asks to REMOVE)
    ws.print_area = 'A1:G20'

    # Leave default margins (openpyxl defaults are ~0.75" all around)
    # Leave default scaling

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
