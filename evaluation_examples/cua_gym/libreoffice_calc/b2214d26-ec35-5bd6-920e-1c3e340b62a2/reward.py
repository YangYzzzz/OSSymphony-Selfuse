"""
Reward Script: Configure print settings for mailing labels on 'Labels' sheet
Task ID: calc_mcp_090
Domain: libreoffice_calc
Scoring:
  Component 1: Top/Bottom margins set to 0.5" (0.3 pts)
  Component 2: Left/Right margins set to 0.19" (0.3 pts)
  Component 3: Scale=100% with no fit-to-page scaling (0.2 pts)
  Component 4: Print area removed (0.2 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_090'

# Margin tolerance: 0.02 inches to account for float rounding
MARGIN_TOL = 0.02


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Labels' sheet must exist
    if 'Labels' not in wb.sheetnames:
        print(f"FAIL: 'Labels' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Labels']
    pm = ws.page_margins
    ps = ws.page_setup

    # Component 1: Top and Bottom margins = 0.5" each (0.3 points)
    # Initial state: top=1.0, bottom=1.0 -> must change to 0.5
    try:
        top_ok = abs(pm.top - 0.5) <= MARGIN_TOL
        bottom_ok = abs(pm.bottom - 0.5) <= MARGIN_TOL
        if top_ok and bottom_ok:
            print(f"PASS: Component 1 — Top margin={pm.top}, Bottom margin={pm.bottom} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Top margin={pm.top} (expected 0.5), Bottom margin={pm.bottom} (expected 0.5)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Left and Right margins = 0.19" each (0.3 points)
    # Initial state: left=0.75, right=0.75 -> must change to 0.19
    try:
        left_ok = abs(pm.left - 0.19) <= MARGIN_TOL
        right_ok = abs(pm.right - 0.19) <= MARGIN_TOL
        if left_ok and right_ok:
            print(f"PASS: Component 2 — Left margin={pm.left}, Right margin={pm.right} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Left margin={pm.left} (expected 0.19), Right margin={pm.right} (expected 0.19)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Scale set to 100% with no fit-to-page (0.2 points)
    # Initial state: scale=100 BUT fitToWidth=1, fitToHeight=1 (fit-to-page enabled)
    # Golden state: scale=100 AND fitToWidth=None, fitToHeight=None (pure 100% scale)
    # We verify that fit-to-page scaling is disabled (the actual task-introduced change)
    try:
        scale_val = ps.scale
        ftw = ps.fitToWidth
        fth = ps.fitToHeight
        # Scale should be 100 AND fitToWidth/fitToHeight should be disabled (None or 0)
        scale_ok = (scale_val is not None and int(scale_val) == 100)
        fit_disabled = (ftw is None or int(ftw) == 0) and (fth is None or int(fth) == 0)
        if scale_ok and fit_disabled:
            print(f"PASS: Component 3 — Scale={scale_val}, fitToWidth={ftw}, fitToHeight={fth} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Scale={scale_val}, fitToWidth={ftw}, fitToHeight={fth} (expected scale=100, no fit-to-page)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Print area removed — no restrictions (0.2 points)
    # Initial state: print_area='Labels'!$A$1:$G$20 -> must be removed
    try:
        print_area = ws.print_area
        # print_area should be empty/None/falsy when no print area is set
        if not print_area or (isinstance(print_area, (list, str)) and len(print_area) == 0):
            print(f"PASS: Component 4 — Print area removed (no restriction) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Print area still set: {print_area}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
