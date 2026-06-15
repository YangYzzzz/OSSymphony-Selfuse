"""
Reward Script: Dynamic commission calculator with tiered rates, quarterly bonuses, and annual accelerators
Task ID: calc_sales_052
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): F column — Annual Sales = SUM(B:E) formulas
  Component 2 (0.20): I column — Base Commission = F*0.07 formulas
  Component 3 (0.25): J column — Q Bonuses = count quarters exceeding quota * 2000
  Component 4 (0.15): K column — Annual Multiplier = IF(F/H>1.2, 1.5, 1)
  Component 5 (0.15): L column — Total Compensation = (I+J)*K
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_052'

def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')

def is_sum_formula(val, row):
    """Check if val is a SUM formula covering B:E for the given row."""
    norm = normalize_formula(val)
    # Accept =SUM(B<row>:E<row>) or equivalent
    expected = f'=SUM(B{row}:E{row})'
    if norm == normalize_formula(expected):
        return True
    # Also accept =B<row>+C<row>+D<row>+E<row>
    alt = f'=B{row}+C{row}+D{row}+E{row}'
    if norm == normalize_formula(alt):
        return True
    return False

def is_base_commission_formula(val, row):
    """Check if val computes F*0.07 for the given row."""
    norm = normalize_formula(val)
    # Accept =F<row>*0.07 or =0.07*F<row>
    patterns = [
        f'=F{row}*0.07',
        f'=0.07*F{row}',
        f'=F{row}*7%',
        f'=F{row}*7/100',
    ]
    return any(norm == normalize_formula(p) for p in patterns)

def is_q_bonuses_formula(val, row):
    """Check if val counts quarters exceeding quota and multiplies by 2000."""
    norm = normalize_formula(val)
    if not norm.startswith('='):
        return False
    # The golden uses: =((B>G)+(C>G)+(D>G)+(E>G))*2000
    # Accept COUNTIF-based or arithmetic sum of boolean comparisons
    r = str(row)
    # Pattern: =((B<r>>G<r>)+(C<r>>G<r>)+(D<r>>G<r>)+(E<r>>G<r>))*2000
    expected_pattern = f'=((B{r}>G{r})+(C{r}>G{r})+(D{r}>G{r})+(E{r}>G{r}))*2000'
    if norm == normalize_formula(expected_pattern):
        return True
    # Also accept COUNTIF variant
    # =COUNTIF(B<r>:E<r>,">"&G<r>)*2000
    countif_pattern = f'=COUNTIF(B{r}:E{r},">"&G{r})*2000'
    if norm == normalize_formula(countif_pattern):
        return True
    # Also accept SUMPRODUCT variant
    # =SUMPRODUCT((B<r>:E<r>>G<r>)*1)*2000
    # Be flexible: accept if it references B,C,D,E, G and 2000
    if '2000' in norm and f'G{r}' in norm:
        # Check that it references all four quarters
        has_quarters = all(f'{col}{r}' in norm for col in ['B', 'C', 'D', 'E'])
        if has_quarters:
            return True
    return False

def is_annual_multiplier_formula(val, row):
    """Check if val is IF(F/H>1.2, 1.5, 1)."""
    norm = normalize_formula(val)
    r = str(row)
    # Accept =IF(F<r>/H<r>>1.2,1.5,1)
    patterns = [
        f'=IF(F{r}/H{r}>1.2,1.5,1)',
        f'=IF(F{r}/H{r}>120%,1.5,1)',
        f'=IF((F{r}/H{r})>1.2,1.5,1)',
    ]
    return any(norm == normalize_formula(p) for p in patterns)

def is_total_comp_formula(val, row):
    """Check if val is (I+J)*K."""
    norm = normalize_formula(val)
    r = str(row)
    patterns = [
        f'=(I{r}+J{r})*K{r}',
        f'=K{r}*(I{r}+J{r})',
        f'=(J{r}+I{r})*K{r}',
        f'=K{r}*(J{r}+I{r})',
    ]
    return any(norm == normalize_formula(p) for p in patterns)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that CommCalc sheet exists
    if 'CommCalc' not in wb.sheetnames:
        print("FAIL: Sheet 'CommCalc' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CommCalc']

    # Determine data rows (rows with rep names in column A, starting from row 2)
    data_rows = []
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is not None:
            data_rows.append(r)

    if not data_rows:
        print("FAIL: No data rows found")
        print("REWARD: 0.0")
        return 0.0

    num_rows = len(data_rows)
    print(f"INFO: Found {num_rows} data rows: {data_rows}")

    # Component 1: F column — Annual Sales = SUM(B:E) (0.25 points)
    try:
        f_pass = 0
        for r in data_rows:
            val = ws.cell(row=r, column=6).value  # F column
            if is_sum_formula(val, r):
                f_pass += 1
            else:
                print(f"  DETAIL: F{r} = {val!r} — not a valid SUM formula")
        f_ratio = f_pass / num_rows
        f_score = 0.25 * f_ratio
        if f_pass == num_rows:
            print(f"PASS: Component 1 — F column SUM formulas ({f_pass}/{num_rows}) (0.25 pts)")
            total_score += 0.25
        elif f_pass > 0:
            print(f"PARTIAL: Component 1 — F column SUM formulas ({f_pass}/{num_rows}) ({f_score:.3f} pts)")
            total_score += f_score
        else:
            print(f"FAIL: Component 1 — No valid F column SUM formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: I column — Base Commission = F*0.07 (0.20 points)
    try:
        i_pass = 0
        for r in data_rows:
            val = ws.cell(row=r, column=9).value  # I column
            if is_base_commission_formula(val, r):
                i_pass += 1
            else:
                print(f"  DETAIL: I{r} = {val!r} — not a valid commission formula")
        i_ratio = i_pass / num_rows
        i_score = 0.20 * i_ratio
        if i_pass == num_rows:
            print(f"PASS: Component 2 — I column base commission formulas ({i_pass}/{num_rows}) (0.20 pts)")
            total_score += 0.20
        elif i_pass > 0:
            print(f"PARTIAL: Component 2 — I column base commission formulas ({i_pass}/{num_rows}) ({i_score:.3f} pts)")
            total_score += i_score
        else:
            print(f"FAIL: Component 2 — No valid I column commission formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: J column — Q Bonuses = count quarters exceeding quota * 2000 (0.25 points)
    try:
        j_pass = 0
        for r in data_rows:
            val = ws.cell(row=r, column=10).value  # J column
            if is_q_bonuses_formula(val, r):
                j_pass += 1
            else:
                print(f"  DETAIL: J{r} = {val!r} — not a valid Q bonuses formula")
        j_ratio = j_pass / num_rows
        j_score = 0.25 * j_ratio
        if j_pass == num_rows:
            print(f"PASS: Component 3 — J column Q bonuses formulas ({j_pass}/{num_rows}) (0.25 pts)")
            total_score += 0.25
        elif j_pass > 0:
            print(f"PARTIAL: Component 3 — J column Q bonuses formulas ({j_pass}/{num_rows}) ({j_score:.3f} pts)")
            total_score += j_score
        else:
            print(f"FAIL: Component 3 — No valid J column Q bonuses formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: K column — Annual Multiplier = IF(F/H>1.2, 1.5, 1) (0.15 points)
    try:
        k_pass = 0
        for r in data_rows:
            val = ws.cell(row=r, column=11).value  # K column
            if is_annual_multiplier_formula(val, r):
                k_pass += 1
            else:
                print(f"  DETAIL: K{r} = {val!r} — not a valid multiplier formula")
        k_ratio = k_pass / num_rows
        k_score = 0.15 * k_ratio
        if k_pass == num_rows:
            print(f"PASS: Component 4 — K column annual multiplier formulas ({k_pass}/{num_rows}) (0.15 pts)")
            total_score += 0.15
        elif k_pass > 0:
            print(f"PARTIAL: Component 4 — K column annual multiplier formulas ({k_pass}/{num_rows}) ({k_score:.3f} pts)")
            total_score += k_score
        else:
            print(f"FAIL: Component 4 — No valid K column multiplier formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: L column — Total Compensation = (I+J)*K (0.15 points)
    try:
        l_pass = 0
        for r in data_rows:
            val = ws.cell(row=r, column=12).value  # L column
            if is_total_comp_formula(val, r):
                l_pass += 1
            else:
                print(f"  DETAIL: L{r} = {val!r} — not a valid total comp formula")
        l_ratio = l_pass / num_rows
        l_score = 0.15 * l_ratio
        if l_pass == num_rows:
            print(f"PASS: Component 5 — L column total compensation formulas ({l_pass}/{num_rows}) (0.15 pts)")
            total_score += 0.15
        elif l_pass > 0:
            print(f"PARTIAL: Component 5 — L column total compensation formulas ({l_pass}/{num_rows}) ({l_score:.3f} pts)")
            total_score += l_score
        else:
            print(f"FAIL: Component 5 — No valid L column total comp formulas found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")

# Main entry
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
