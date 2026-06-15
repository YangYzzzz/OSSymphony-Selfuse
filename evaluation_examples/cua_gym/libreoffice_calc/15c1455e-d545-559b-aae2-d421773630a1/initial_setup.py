"""
Initial Setup: Commission Calculator with tiered rates, quarterly bonuses, and annual accelerators
Task ID: calc_sales_052
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CommCalc ---
    ws = wb.active
    ws.title = 'CommCalc'

    # Headers
    headers = [
        'Rep', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales',
        'Annual Sales', 'Quarterly Quota', 'Annual Quota',
        'Base Commission', 'Q Bonuses', 'Annual Multiplier', 'Total Compensation'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - sales reps with quarterly data, quotas
    # Columns: A=Rep, B=Q1, C=Q2, D=Q3, E=Q4, F=Annual(empty), G=QQuota, H=AnnualQuota, I-L=empty
    data = [
        # Rep, Q1 Sales, Q2 Sales, Q3 Sales, Q4 Sales, Annual Sales, Q Quota, Ann Quota
        ['Alice', 80000, 95000, 88000, 110000, None, 75000, 300000, None, None, None, None],
        ['Bob', 70000, 65000, 90000, 85000, None, 75000, 300000, None, None, None, None],
        ['Carol', 92000, 87000, 78000, 95000, None, 85000, 340000, None, None, None, None],
        ['David', 60000, 72000, 68000, 55000, None, 75000, 300000, None, None, None, None],
        ['Elena', 105000, 98000, 112000, 120000, None, 90000, 360000, None, None, None, None],
        ['Frank', 45000, 52000, 48000, 61000, None, 60000, 240000, None, None, None, None],
        ['Grace', 88000, 91000, 85000, 93000, None, 80000, 320000, None, None, None, None],
        ['Henry', 75000, 68000, 82000, 79000, None, 75000, 300000, None, None, None, None],
        ['Iris', 110000, 115000, 108000, 125000, None, 100000, 400000, None, None, None, None],
        ['Jack', 55000, 62000, 58000, 71000, None, 65000, 260000, None, None, None, None],
    ]

    data_font = Font(name='Calibri', size=11)
    currency_format = '$#,##0'
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if val is not None:
                    cell.number_format = currency_format

    # Set column widths
    col_widths = {
        'A': 14, 'B': 14, 'C': 14, 'D': 14, 'E': 14,
        'F': 16, 'G': 18, 'H': 16,
        'I': 18, 'J': 14, 'K': 18, 'L': 22,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet: Instructions ---
    ws2 = wb.create_sheet('Instructions')
    ws2['A1'] = 'Commission Calculator Instructions'
    ws2['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws2['A3'] = 'Base Commission Rate: 7% of Annual Sales'
    ws2['A4'] = 'Quarterly Bonus: $2,000 for each quarter where sales exceed the Quarterly Quota'
    ws2['A5'] = 'Annual Accelerator: 1.5x multiplier if Annual Attainment > 120%'
    ws2['A6'] = 'Total Compensation = (Base Commission + Q Bonuses) * Annual Multiplier'
    ws2['A8'] = 'Formulas needed in CommCalc sheet:'
    ws2['A9'] = '  F: Annual Sales = SUM of Q1 through Q4'
    ws2['A10'] = '  I: Base Commission = Annual Sales * 0.07'
    ws2['A11'] = '  J: Q Bonuses = Count of quarters exceeding quota * $2,000'
    ws2['A12'] = '  K: Annual Multiplier = 1.5 if (Annual Sales / Annual Quota) > 1.2, else 1'
    ws2['A13'] = '  L: Total Compensation = (Base Commission + Q Bonuses) * Annual Multiplier'
    ws2.column_dimensions['A'].width = 70

    for row_num in range(3, 14):
        ws2.cell(row=row_num, column=1).font = Font(name='Calibri', size=11)

    # --- Sheet: QuotaReference ---
    ws3 = wb.create_sheet('QuotaReference')
    ws3['A1'] = 'Department'
    ws3['B1'] = 'Standard Quarterly Quota'
    ws3['C1'] = 'Standard Annual Quota'
    ws3['D1'] = 'Notes'
    for c in range(1, 5):
        cell = ws3.cell(row=1, column=c)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    quota_data = [
        ['Enterprise Sales', 90000, 360000, 'Top-tier accounts'],
        ['Mid-Market', 75000, 300000, 'Standard quota'],
        ['SMB', 60000, 240000, 'Volume-focused'],
        ['Strategic', 100000, 400000, 'Key accounts only'],
        ['Channel Partners', 80000, 320000, 'Partner-driven'],
        ['Inside Sales', 65000, 260000, 'Inbound leads'],
    ]
    for r, row_data in enumerate(quota_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 24
    ws3.column_dimensions['C'].width = 22
    ws3.column_dimensions['D'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
