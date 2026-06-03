"""
Reward Script: Build a professional invoice template for a consulting firm
Task ID: calc_fin_invoice_template_016
Domain: libreoffice_calc
Scoring:
  Component 1: Header structure (merged cells, text, bold, font sizes, alignment)      0.25
  Component 2: Client info area (labels, merged cells, light yellow fills)              0.20
  Component 3: Line items (row 9 headers bold, E10:E19 IF formulas, A10:D19 yellow)    0.25
  Component 4: Totals area (E21 SUM, E22 tax 8.5%, E23 grand total, currency format)   0.15
  Component 5: Sheet protection enabled                                                  0.15
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_invoice_template_016'
YELLOW_RGB = 'FFFFFF99'  # Light yellow fill used in golden file


def get_merged_ranges(ws):
    """Return set of merged range strings."""
    return set(str(r) for r in ws.merged_cells.ranges)


def check_cell_yellow(ws, coord):
    """Check if a cell has light yellow fill (FFFFFF99)."""
    try:
        cell = ws[coord]
        return (cell.fill.fill_type == 'solid' and
                cell.fill.fgColor.rgb == YELLOW_RGB)
    except Exception:
        return False


def check_formula_contains(ws, coord, fragment):
    """Check if cell formula contains fragment (case-insensitive, no spaces)."""
    try:
        val = ws[coord].value
        if not isinstance(val, str):
            return False
        return fragment.upper().replace(' ', '') in val.upper().replace(' ', '')
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet 'Invoice' exists
    if 'Invoice' not in wb.sheetnames:
        print("FAIL: Sheet 'Invoice' not found")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoice']
    merged_ranges = get_merged_ranges(ws)

    # ------------------------------------------------------------------
    # Component 1: Header structure (0.25 points)
    # Checks:
    #   - A1:F1 merged, value='INVOICE', bold, font size 18, centered
    #   - A2:F2 merged, value='Acme Consulting LLC', font size 12, centered
    # All of these are absent in the initial (blank) file.
    # ------------------------------------------------------------------
    try:
        comp1_score = 0.0

        # 1a: A1:F1 merged
        a1_merged = 'A1:F1' in merged_ranges
        # 1b: A1 value is 'INVOICE'
        a1_val = ws['A1'].value
        a1_text = isinstance(a1_val, str) and a1_val.strip().upper() == 'INVOICE'
        # 1c: A1 bold
        a1_bold = ws['A1'].font.bold is True
        # 1d: A1 font size >= 16 (task says 18)
        a1_size = ws['A1'].font.size is not None and ws['A1'].font.size >= 16
        # 1e: A1 centered
        a1_center = ws['A1'].alignment.horizontal == 'center'

        # 1f: A2:F2 merged
        a2_merged = 'A2:F2' in merged_ranges
        # 1g: A2 value contains company name
        a2_val = ws['A2'].value
        a2_text = isinstance(a2_val, str) and 'Acme Consulting' in a2_val
        # 1h: A2 font size >= 11
        a2_size = ws['A2'].font.size is not None and ws['A2'].font.size >= 11
        # 1i: A2 centered
        a2_center = ws['A2'].alignment.horizontal == 'center'

        header_checks = [a1_merged, a1_text, a1_bold, a1_size, a1_center,
                         a2_merged, a2_text, a2_size, a2_center]
        passed = sum(1 for c in header_checks if c)

        # Award full points if majority pass; partial otherwise
        if passed >= 8:
            comp1_score = 0.25
            print(f"PASS: Component 1 — Header structure fully correct ({passed}/9 checks) (0.25 pts)")
        elif passed >= 5:
            comp1_score = 0.15
            print(f"PARTIAL: Component 1 — Header partially correct ({passed}/9 checks) (0.15 pts)")
        elif passed >= 3:
            comp1_score = 0.08
            print(f"PARTIAL: Component 1 — Header minimally correct ({passed}/9 checks) (0.08 pts)")
        else:
            print(f"FAIL: Component 1 — Header structure missing or wrong ({passed}/9 checks)")
            print(f"  a1_merged={a1_merged}, a1_text={a1_text}, a1_bold={a1_bold}, "
                  f"a1_size={a1_size}, a1_center={a1_center}")
            print(f"  a2_merged={a2_merged}, a2_text={a2_text}, a2_size={a2_size}, a2_center={a2_center}")

        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Client info area (0.20 points)
    # Checks:
    #   - A4='Bill To:', A5='Address:', A6='Invoice Date:', A7='Invoice #:'
    #   - B4:C4 merged with yellow fill; B5:C5 merged with yellow fill
    #   - B6 yellow fill; B7 yellow fill
    # None of these exist in the initial (blank) file.
    # ------------------------------------------------------------------
    try:
        comp2_score = 0.0

        # Labels
        a4_label = isinstance(ws['A4'].value, str) and 'Bill' in ws['A4'].value
        a5_label = isinstance(ws['A5'].value, str) and 'Address' in ws['A5'].value
        a6_label = isinstance(ws['A6'].value, str) and 'Invoice Date' in ws['A6'].value
        a7_label = isinstance(ws['A7'].value, str) and 'Invoice' in ws['A7'].value and '#' in ws['A7'].value

        # Merged yellow ranges for client details
        b4c4_merged = 'B4:C4' in merged_ranges
        b5c5_merged = 'B5:C5' in merged_ranges

        # Yellow fills on entry cells
        b4_yellow = check_cell_yellow(ws, 'B4')
        b5_yellow = check_cell_yellow(ws, 'B5')
        b6_yellow = check_cell_yellow(ws, 'B6')
        b7_yellow = check_cell_yellow(ws, 'B7')

        info_checks = [a4_label, a5_label, a6_label, a7_label,
                       b4c4_merged, b5c5_merged,
                       b4_yellow, b5_yellow, b6_yellow, b7_yellow]
        passed = sum(1 for c in info_checks if c)

        if passed >= 8:
            comp2_score = 0.20
            print(f"PASS: Component 2 — Client info area correct ({passed}/10 checks) (0.20 pts)")
        elif passed >= 5:
            comp2_score = 0.12
            print(f"PARTIAL: Component 2 — Client info area partially correct ({passed}/10 checks) (0.12 pts)")
        elif passed >= 2:
            comp2_score = 0.06
            print(f"PARTIAL: Component 2 — Client info area minimally correct ({passed}/10 checks) (0.06 pts)")
        else:
            print(f"FAIL: Component 2 — Client info area missing ({passed}/10 checks)")
            print(f"  labels: a4={a4_label}, a5={a5_label}, a6={a6_label}, a7={a7_label}")
            print(f"  merges: b4c4={b4c4_merged}, b5c5={b5c5_merged}")
            print(f"  yellow: b4={b4_yellow}, b5={b5_yellow}, b6={b6_yellow}, b7={b7_yellow}")

        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Line items area (0.25 points)
    # Checks:
    #   - Row 9 headers: Description, Qty, Unit Price, Discount %, Line Total — bold
    #   - E10:E19 each has IF formula: =IF(Ax="","",Bx*Cx*(1-Dx))
    #   - A10:D19 cells have light yellow fill (40 cells total, check sample)
    # None of these exist in the initial (blank) file.
    # ------------------------------------------------------------------
    try:
        comp3_score = 0.0

        # Row 9 headers bold
        headers_expected = {1: 'Description', 2: 'Qty', 3: 'Unit Price',
                            4: 'Discount', 5: 'Line Total'}
        headers_ok = 0
        for col, expected_fragment in headers_expected.items():
            cell = ws.cell(row=9, column=col)
            val = cell.value
            if (isinstance(val, str) and expected_fragment.lower() in val.lower()
                    and cell.font.bold is True):
                headers_ok += 1

        # E10:E19 IF formulas
        if_formula_count = 0
        for row in range(10, 20):
            cell = ws.cell(row=row, column=5)
            val = cell.value
            if isinstance(val, str) and val.upper().startswith('=IF('):
                if_formula_count += 1

        # Yellow fill on A10:D19 (check key cells)
        yellow_sample = []
        for row in [10, 12, 15, 19]:
            for col in [1, 2, 3, 4]:
                from openpyxl.utils import get_column_letter
                coord = f'{get_column_letter(col)}{row}'
                yellow_sample.append(check_cell_yellow(ws, coord))
        yellow_ok = sum(1 for y in yellow_sample if y)

        # Scoring: 3 sub-checks
        headers_pass = headers_ok >= 4
        formulas_pass = if_formula_count >= 8
        yellow_pass = yellow_ok >= 12

        passed_sub = sum([headers_pass, formulas_pass, yellow_pass])

        if passed_sub == 3:
            comp3_score = 0.25
            print(f"PASS: Component 3 — Line items complete: headers={headers_ok}/5, "
                  f"IF formulas={if_formula_count}/10, yellow cells={yellow_ok}/16 (0.25 pts)")
        elif passed_sub == 2:
            comp3_score = 0.16
            print(f"PARTIAL: Component 3 — Line items partially complete ({passed_sub}/3 sub-checks): "
                  f"headers={headers_ok}/5, IF formulas={if_formula_count}/10, yellow={yellow_ok}/16 (0.16 pts)")
        elif passed_sub == 1:
            comp3_score = 0.08
            print(f"PARTIAL: Component 3 — Line items minimally complete ({passed_sub}/3 sub-checks) (0.08 pts)")
        else:
            print(f"FAIL: Component 3 — Line items area missing: "
                  f"headers={headers_ok}/5 (pass>={4}), "
                  f"IF formulas={if_formula_count}/10 (pass>={8}), "
                  f"yellow={yellow_ok}/16 (pass>={12})")

        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: Totals area formulas and currency format (0.15 points)
    # Checks:
    #   - A21='Subtotal', E21 contains =SUM(E10:E19) formula
    #   - A22='Tax (8.5%)', E22 contains tax formula referencing 0.085
    #   - A23='TOTAL', E23 contains sum formula =E21+E22
    #   - E21:E23 have currency number format ($#,##0.00)
    # None exist in initial blank file.
    # ------------------------------------------------------------------
    try:
        comp4_score = 0.0

        # Subtotal row
        a21_text = isinstance(ws['A21'].value, str) and 'subtotal' in ws['A21'].value.lower()
        e21_formula = check_formula_contains(ws, 'E21', 'SUM(E10:E19)')

        # Tax row
        a22_text = isinstance(ws['A22'].value, str) and 'tax' in ws['A22'].value.lower()
        e22_formula = (isinstance(ws['E22'].value, str) and
                       '0.085' in ws['E22'].value)

        # Total row
        a23_text = isinstance(ws['A23'].value, str) and 'total' in ws['A23'].value.upper()
        e23_formula = (isinstance(ws['E23'].value, str) and
                       'E21' in ws['E23'].value.upper() and
                       'E22' in ws['E23'].value.upper())

        # Currency format
        currency_format = '$#,##0.00'
        e21_currency = currency_format in ws['E21'].number_format
        e22_currency = currency_format in ws['E22'].number_format
        e23_currency = currency_format in ws['E23'].number_format

        totals_checks = [a21_text, e21_formula, a22_text, e22_formula,
                         a23_text, e23_formula, e21_currency, e22_currency, e23_currency]
        passed = sum(1 for c in totals_checks if c)

        if passed >= 7:
            comp4_score = 0.15
            print(f"PASS: Component 4 — Totals area formulas correct ({passed}/9 checks) (0.15 pts)")
        elif passed >= 4:
            comp4_score = 0.09
            print(f"PARTIAL: Component 4 — Totals area partially correct ({passed}/9 checks) (0.09 pts)")
        elif passed >= 2:
            comp4_score = 0.05
            print(f"PARTIAL: Component 4 — Totals area minimally correct ({passed}/9 checks) (0.05 pts)")
        else:
            print(f"FAIL: Component 4 — Totals area missing or wrong ({passed}/9 checks)")
            print(f"  a21={a21_text}, e21_formula={e21_formula}, a22={a22_text}, "
                  f"e22={e22_formula}, a23={a23_text}, e23={e23_formula}")
            print(f"  currency: e21={e21_currency}, e22={e22_currency}, e23={e23_currency}")

        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ------------------------------------------------------------------
    # Component 5: Sheet protection enabled (0.15 points)
    # Checks:
    #   - ws.protection.sheet is True (sheet protection enabled)
    # The key task-introduced change is enabling sheet protection.
    # Note: cells are locked=True by default in xlsx, so locking alone is
    # not a distinguishing marker — only the ACTIVATION of sheet protection
    # (protection.sheet=True) separates initial from golden.
    # Initial file has protection.sheet=False; golden has protection.sheet=True.
    # ------------------------------------------------------------------
    try:
        comp5_score = 0.0

        sheet_protected = ws.protection.sheet is True

        if sheet_protected:
            comp5_score = 0.15
            print(f"PASS: Component 5 — Sheet protection is enabled (0.15 pts)")
        else:
            print(f"FAIL: Component 5 — Sheet protection not enabled "
                  f"(ws.protection.sheet={ws.protection.sheet})")

        total_score += comp5_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
