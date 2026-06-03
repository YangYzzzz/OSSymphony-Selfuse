"""
Initial Setup: NPS Matrix spreadsheet with regional month-by-month NPS scores
Task ID: calc_gen_chart_051
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_chart_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'NPSMatrix'

    # Row 1: blank A1, then months Jan-Dec in B1:M1
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    ws['A1'] = None
    for col_idx, month in enumerate(months, start=2):
        ws.cell(row=1, column=col_idx, value=month)

    # Column A: regions in A2:A6
    regions = ['North', 'South', 'East', 'West', 'Central']
    for row_idx, region in enumerate(regions, start=2):
        ws.cell(row=row_idx, column=1, value=region)

    # B2:M6: NPS scores (integers -100 to +100), realistic and varied
    # Each row is a region (North, South, East, West, Central)
    # Each column is a month (Jan-Dec)
    nps_data = [
        # North: generally positive trend
        [42, 45, 38, 50, 55, 60, 58, 63, 67, 70, 72, 75],
        # South: moderate, somewhat volatile
        [18, 22, 15, 28, 30, 25, 35, 32, 28, 40, 38, 42],
        # East: struggling, gradual recovery
        [-12, -8, -15, -5, 2, 8, 12, 18, 22, 25, 28, 32],
        # West: strong performer
        [55, 58, 62, 65, 68, 72, 70, 75, 78, 80, 82, 85],
        # Central: flat with slight improvement
        [5, 8, 3, 10, 12, 15, 18, 14, 20, 22, 25, 28],
    ]

    for row_idx, region_scores in enumerate(nps_data, start=2):
        for col_idx, score in enumerate(region_scores, start=2):
            ws.cell(row=row_idx, column=col_idx, value=score)

    # Rows 7 onward: empty (no data, no formatting)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col_letter].width = 8

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
