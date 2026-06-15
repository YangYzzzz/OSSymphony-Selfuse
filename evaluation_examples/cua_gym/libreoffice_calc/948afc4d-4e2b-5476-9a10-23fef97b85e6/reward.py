"""
Reward Script: NPS Heat Map Visualization with Color Scale and Trend
Task ID: calc_gen_chart_051
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Conditional formatting (color scale) on B2:M6 — 0.30 pts
  Component 2: Column N header 'Trend' + REPT trend formulas in N2:N6 — 0.20 pts
  Component 3: Column O header 'Avg NPS' + AVERAGE formulas in O2:O6 — 0.20 pts
  Component 4: Conditional formatting (color scale) on O2:O6 — 0.15 pts
  Component 5: Row 7 'Monthly Avg' label + AVERAGE formulas in B7:M7 — 0.15 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_chart_051'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'NPSMatrix' must exist
    if 'NPSMatrix' not in wb.sheetnames:
        print("CRITICAL: Sheet 'NPSMatrix' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['NPSMatrix']

    # -------------------------------------------------------------------------
    # Component 1: Color scale conditional formatting on B2:M6 (0.30 points)
    # 3-color scale: red at -100, white at 0, green at +100
    # This is absent in the initial file — only present after task completion.
    # -------------------------------------------------------------------------
    try:
        cf_rules_all = list(ws.conditional_formatting)
        # Find a colorScale rule covering B2:M6 (or at least B2:M6 range)
        matrix_cf_found = False
        for cf in cf_rules_all:
            range_str = str(cf)
            if 'B2:M6' in range_str:
                for rule in cf.rules:
                    if rule.type == 'colorScale' and rule.colorScale is not None:
                        cs = rule.colorScale
                        # Verify 3-stop color scale
                        if len(cs.cfvo) == 3 and len(cs.color) == 3:
                            # Check the stops cover red-white-green logic
                            # Stop 0 (min/low): should be red-ish
                            # Stop 1 (mid): should be white-ish
                            # Stop 2 (max/high): should be green-ish
                            c0_rgb = cs.color[0].rgb if cs.color[0].rgb else ''
                            c1_rgb = cs.color[1].rgb if cs.color[1].rgb else ''
                            c2_rgb = cs.color[2].rgb if cs.color[2].rgb else ''

                            # Red channel high in stop 0 (red-ish), RGB: FF or similar
                            # Green channel high in stop 2 (green-ish)
                            # White in stop 1 (FFFFFF or similar)
                            # We check that the colors are meaningfully different
                            colors_valid = (
                                c0_rgb != c2_rgb and  # low and high stops differ
                                len(c0_rgb) >= 6 and
                                len(c1_rgb) >= 6 and
                                len(c2_rgb) >= 6
                            )

                            # Also check the value anchors (num type with -100, 0, 100)
                            # or percentile/min/max type
                            cfvo_types = [v.type for v in cs.cfvo]
                            # Accept 'num' type with appropriate ranges OR min/max/percentile
                            anchors_valid = False
                            if all(t == 'num' for t in cfvo_types):
                                vals = [v.val for v in cs.cfvo]
                                # Check values are ordered (ascending) with range covering negatives and positives
                                if (vals[0] is not None and vals[2] is not None and
                                        float(vals[0]) < 0 and float(vals[2]) > 0 and
                                        float(vals[0]) <= float(vals[1]) <= float(vals[2])):
                                    anchors_valid = True
                            elif cfvo_types[0] in ('min', 'percentile') or cfvo_types[2] in ('max', 'percentile'):
                                anchors_valid = True

                            if colors_valid and anchors_valid:
                                matrix_cf_found = True
                                print(f"PASS: Component 1 — Color scale CF on B2:M6 found with 3-stop red-white-green scale (0.30 pts)")
                                print(f"       Stops: {c0_rgb} | {c1_rgb} | {c2_rgb}")
                                total_score += 0.30
                                break
                        if matrix_cf_found:
                            break
            if matrix_cf_found:
                break

        if not matrix_cf_found:
            print(f"FAIL: Component 1 — No valid 3-color scale CF found on B2:M6")
            print(f"       CF rules found: {[str(cf) for cf in cf_rules_all]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Trend column — N1='Trend' header, N2:N6 REPT-based formulas (0.20 points)
    # -------------------------------------------------------------------------
    try:
        n1_val = ws['N1'].value
        header_ok = (n1_val is not None and str(n1_val).strip().lower() == 'trend')

        # Check N2:N6 have formulas (REPT or similar sparkline-style formula)
        formula_count = 0
        for row_idx in range(2, 7):
            cell_val = ws.cell(row=row_idx, column=14).value  # Column N
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                formula_count += 1

        if header_ok and formula_count == 5:
            print(f"PASS: Component 2 — N1='Trend' header present and N2:N6 contain formulas ({formula_count}/5 cells) (0.20 pts)")
            total_score += 0.20
        elif header_ok and formula_count > 0:
            partial = round(0.10 * formula_count / 5, 2)
            print(f"PARTIAL: Component 2 — N1='Trend' header OK, but only {formula_count}/5 trend formulas in N2:N6")
            total_score += 0.10 + partial
        elif header_ok:
            print(f"PARTIAL: Component 2 — N1='Trend' header present but no formulas in N2:N6")
            total_score += 0.05
        else:
            found_header = repr(n1_val)
            print(f"FAIL: Component 2 — Expected N1='Trend', found {found_header}; formula_count={formula_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Avg NPS column — O1='Avg NPS' header, O2:O6 AVERAGE formulas (0.20 points)
    # -------------------------------------------------------------------------
    try:
        o1_val = ws['O1'].value
        header_ok = (o1_val is not None and 'avg' in str(o1_val).strip().lower())

        # Check O2:O6 have AVERAGE formulas
        avg_formula_count = 0
        for row_idx in range(2, 7):
            cell_val = ws.cell(row=row_idx, column=15).value  # Column O
            if (cell_val is not None and isinstance(cell_val, str) and
                    cell_val.startswith('=') and 'AVERAGE' in cell_val.upper()):
                avg_formula_count += 1

        if header_ok and avg_formula_count == 5:
            print(f"PASS: Component 3 — O1 header present ('{o1_val}') and O2:O6 AVERAGE formulas ({avg_formula_count}/5) (0.20 pts)")
            total_score += 0.20
        elif header_ok and avg_formula_count > 0:
            partial = round(0.10 * avg_formula_count / 5, 2)
            print(f"PARTIAL: Component 3 — O1 header OK, but only {avg_formula_count}/5 AVERAGE formulas in O2:O6")
            total_score += 0.10 + partial
        elif avg_formula_count == 5:
            print(f"PARTIAL: Component 3 — AVERAGE formulas in O2:O6 present but O1 header missing/incorrect (found: {repr(o1_val)})")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — O1={repr(o1_val)}, AVERAGE formulas found: {avg_formula_count}/5")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Color scale conditional formatting on O2:O6 (0.15 points)
    # Same red-white-green scale as the matrix
    # -------------------------------------------------------------------------
    try:
        o_cf_found = False
        for cf in ws.conditional_formatting:
            range_str = str(cf)
            if 'O2:O6' in range_str or 'O2' in range_str:
                for rule in cf.rules:
                    if rule.type == 'colorScale' and rule.colorScale is not None:
                        cs = rule.colorScale
                        if len(cs.cfvo) >= 2 and len(cs.color) >= 2:
                            o_cf_found = True
                            print(f"PASS: Component 4 — Color scale CF on O2:O6 found (0.15 pts)")
                            total_score += 0.15
                            break
                if o_cf_found:
                    break

        if not o_cf_found:
            print(f"FAIL: Component 4 — No color scale CF found on O2:O6")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Row 7 'Monthly Avg' label + AVERAGE formulas B7:M7 (0.15 points)
    # -------------------------------------------------------------------------
    try:
        a7_val = ws['A7'].value
        label_ok = (a7_val is not None and 'avg' in str(a7_val).strip().lower())

        # Check B7:M7 have AVERAGE formulas (columns 2-13)
        monthly_avg_count = 0
        for col_idx in range(2, 14):  # B=2 to M=13
            cell_val = ws.cell(row=7, column=col_idx).value
            if (cell_val is not None and isinstance(cell_val, str) and
                    cell_val.startswith('=') and 'AVERAGE' in cell_val.upper()):
                monthly_avg_count += 1

        if label_ok and monthly_avg_count == 12:
            print(f"PASS: Component 5 — A7='{a7_val}' label and B7:M7 AVERAGE formulas ({monthly_avg_count}/12) (0.15 pts)")
            total_score += 0.15
        elif label_ok and monthly_avg_count > 0:
            partial = round(0.075 * monthly_avg_count / 12, 3)
            print(f"PARTIAL: Component 5 — A7 label OK, only {monthly_avg_count}/12 AVERAGE formulas in B7:M7")
            total_score += 0.075 + partial
        elif monthly_avg_count == 12:
            print(f"PARTIAL: Component 5 — B7:M7 AVERAGE formulas present but A7 label missing/incorrect (found: {repr(a7_val)})")
            total_score += 0.075
        else:
            print(f"FAIL: Component 5 — A7={repr(a7_val)}, AVERAGE formulas found: {monthly_avg_count}/12")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
