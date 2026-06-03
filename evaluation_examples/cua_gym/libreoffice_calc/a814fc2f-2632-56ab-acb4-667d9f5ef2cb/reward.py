"""
FINAL REWARD SCRIPT - SUCCESS
Task: Enable data validation for the "Risk Level" column allowing users to choose from "Critical", "High", "Medium", or "Low" via dropdown list. Finish without modifying unrelated cells.
Generated: 2025-11-24 07:51:23
Status: success
Model: o3
Total Steps: 3
"""

import openpyxl
import re
import os


def verify_task(file_path: str) -> float:
    """Verify that the file contains a list-type data-validation on the
    'Risk Level' column (column C) with the exact options
    Critical, High, Medium, Low and that no original data was modified.

    Returns a progressive score between 0.0 and 1.0 (float).
    """

    print(f"Verifying task for file: {file_path}")
    max_score = 1.0
    total_score = 0.0

    # Expected unchanged initial data (first 4 rows)
    expected_rows = [
        ['ID', 'Description', 'Risk Level', 'Owner'],
        ['1', 'Server outage', '', 'Alice'],
        ['2', 'Data breach', '', 'Bob'],
        ['3', 'Regulatory non-compliance', '', 'Charlie']
    ]

    # ------------------------------------------------------------------
    # 1. Load workbook & access sheet
    # ------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(file_path)
        print("✓ Workbook loaded")
    except Exception as e:
        print(f"✗ Unable to load workbook: {e}")
        return 0.0

    if 'Project Risks' not in wb.sheetnames:
        print("✗ Sheet 'Project Risks' not found")
        return 0.0
    sheet = wb['Project Risks']

    # ------------------------------------------------------------------
    # 2. Verify original data remains unchanged  (0.2 points)
    # ------------------------------------------------------------------
    data_unchanged = True
    for i, expected_row in enumerate(expected_rows, start=1):
        actual_row = [sheet.cell(row=i, column=j+1).value for j in range(len(expected_row))]
        processed_actual = ["" if v is None else str(v) for v in actual_row]
        processed_expected = [str(v) for v in expected_row]
        if processed_actual != processed_expected:
            print(f"✗ Row {i} mismatch. Expected {processed_expected}, found {processed_actual}")
            data_unchanged = False
            break
    if data_unchanged:
        print("✓ Core data rows unchanged (0.2)")
        total_score += 0.2

    # ------------------------------------------------------------------
    # 3. Verify data-validation settings  (up to 0.8 points)
    # ------------------------------------------------------------------
    expected_list = {"critical", "high", "medium", "low"}
    dv_found = None

    # Iterate through every DataValidation rule in the worksheet
    for dv in sheet.data_validations.dataValidation:
        # Check that it is a LIST validation
        if (dv.type or '').lower() != 'list':
            continue

        # ---------------------
        # 3a. Extract the option list
        # ---------------------
        formula = dv.formula1 or ''
        if formula.startswith('='):
            formula_clean = formula[1:]
        else:
            formula_clean = formula
        if formula_clean.startswith('"') and formula_clean.endswith('"'):
            formula_clean = formula_clean[1:-1]
        options = [opt.strip().lower() for opt in formula_clean.split(',') if opt.strip()]
        option_set = set(options)
        if option_set != expected_list:
            continue  # wrong option set, try next rule

        # ---------------------
        # 3b. Verify the rule applies to the Risk Level column and includes C2
        # ---------------------
        sqref_str = str(dv.sqref).replace('$', '').upper()  # handles MultiCellRange -> string
        includes_c2 = False
        for ref in sqref_str.split():  # multiple ranges separated by spaces
            if ':' in ref:
                start_cell, end_cell = ref.split(':')
            else:
                start_cell = end_cell = ref

            def parse(cell):
                m = re.match(r'([A-Z]+)([0-9]*)', cell)
                if m:
                    col = m.group(1)
                    row_str = m.group(2)
                    row = int(row_str) if row_str else None
                    return col, row
                return None, None

            start_col, start_row = parse(start_cell)
            end_col, end_row = parse(end_cell)

            if start_col == 'C' and end_col == 'C':
                # Whole column or explicit range containing row 2
                if start_row is None or end_row is None:
                    includes_c2 = True
                elif start_row <= 2 <= end_row:
                    includes_c2 = True

            if includes_c2:
                break

        if includes_c2:
            dv_found = dv
            break  # perfect rule identified

    # ------------------------------------------------------------------
    # 4. Scoring for data-validation
    # ------------------------------------------------------------------
    if dv_found:
        print("✓ List validation with correct options found")
        print("✓ Validation applies to 'Risk Level' column and includes cell C2")
        total_score += 0.8  # full remaining points for validation correctness
    else:
        # Partial credit: some list validation exists on column C but incorrect
        partial_found = False
        for dv in sheet.data_validations.dataValidation:
            if (dv.type or '').lower() == 'list' and ' C' in str(dv.sqref).upper():
                partial_found = True
                break
        if partial_found:
            print("✓ Some list data validation found in 'Risk Level' column but options or range incorrect (0.3)")
            total_score += 0.3
        else:
            print("✗ No suitable data validation found for 'Risk Level' column")

    # ------------------------------------------------------------------
    # 5. Final score
    # ------------------------------------------------------------------
    final_score = min(total_score, max_score)
    print(f"Final score: {final_score}")
    return final_score


if __name__ == '__main__':
    # Path expected in the task context
    default_path = '/home/user/enable_data_validation_for_the_risk_level_column_allowing_users_to_choose_from_critical_high_medium_.xlsx'
    path_to_check = default_path if os.path.exists(default_path) else ''
    reward = verify_task(path_to_check)
    print(f"REWARD: {reward}")
