"""
FINAL REWARD SCRIPT - SUCCESS
Task: I need the 'Student ID' values copied to 'University ID' with zero-padding on the left to reach 7 digits. Please do this and leave everything else as-is.
Generated: 2025-11-24 07:25:41
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl
import os


def verify_task(file_path: str) -> float:
    """Verify that the Student ID values were copied to University ID with left-side
    zero-padding to 7 digits, while all other columns remain unchanged.

    Scoring (progressive):
        • 0.60 – Every University ID is the zero-padded Student ID
        • 0.20 – All original Student ID values remain unchanged
        • 0.10 – All Names remain unchanged
        • 0.10 – All Majors remain unchanged

    Returns a float between 0.0 and 1.0.
    """
    print(f"Verifying file: {file_path}")

    max_score = 1.0
    weights = {
        "univ_id": 0.60,
        "student_id": 0.20,
        "name": 0.10,
        "major": 0.10,
    }

    # Expected unchanged reference data (taken from the task description)
    reference_rows = [
        ("12345", "Alice Smith", "Physics"),
        ("6789", "Bob Johnson", "Mathematics"),
        ("234567", "Carol White", "Biology"),
        ("89012", "David Brown", "Chemistry"),
        ("3456", "Eve Davis", "History"),
    ]
    reference_lookup = {sid: (name, major) for sid, name, major in reference_rows}

    try:
        wb = openpyxl.load_workbook(file_path)
        print("✓ Workbook loaded")
    except Exception as e:
        print(f"✗ Could not load workbook: {e}")
        return 0.0

    # Prefer sheet named "Students"; fall back to the active sheet
    sheet = wb["Students"] if "Students" in wb.sheetnames else wb.active

    # Validate header row exists and locate column indices
    header_row = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    print("Headers:", header_row)

    try:
        idx_student = header_row.index("Student ID")
        idx_univ = header_row.index("University ID")
        idx_name = header_row.index("Name")
        idx_major = header_row.index("Major")
    except ValueError as err:
        print("✗ Required header missing:", err)
        return 0.0

    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    total = len(data_rows)
    print(f"Found {total} data rows")
    if total == 0:
        print("✗ No data rows present")
        return 0.0

    # Counters for correct entries
    correct_univ = correct_student = correct_name = correct_major = 0

    for row in data_rows:
        student_raw = row[idx_student]
        univ_raw = row[idx_univ]
        name_raw = row[idx_name]
        major_raw = row[idx_major]

        student_str = "" if student_raw is None else str(student_raw).strip()
        univ_str = "" if univ_raw is None else str(univ_raw).strip()

        # Check University ID zero-padding rule
        if univ_str == student_str.zfill(7):
            correct_univ += 1

        # Reference checks to ensure other columns unchanged
        if student_str in reference_lookup:
            correct_student += 1  # Student ID matches one of expected originals
            ref_name, ref_major = reference_lookup[student_str]
            if str(name_raw).strip() == ref_name:
                correct_name += 1
            if str(major_raw).strip() == ref_major:
                correct_major += 1
        else:
            print(f"✗ Unexpected Student ID found: {student_str}")

    # Progressive scoring
    scores = {
        "univ_id": (correct_univ / total) * weights["univ_id"],
        "student_id": (correct_student / total) * weights["student_id"],
        "name": (correct_name / total) * weights["name"],
        "major": (correct_major / total) * weights["major"],
    }

    for k, v in scores.items():
        print(f"{k} score contribution: {v:.2f}")

    final_score = min(sum(scores.values()), max_score)
    print(f"Total score: {final_score}")
    return final_score


if __name__ == "__main__":
    path = "/home/user/i_need_the_student_id_values_copied_to_university_id_with_zero_padding_on_the_left_to_reach_7_digits.xlsx"
    reward = verify_task(path)
    print(f"REWARD: {reward}")
