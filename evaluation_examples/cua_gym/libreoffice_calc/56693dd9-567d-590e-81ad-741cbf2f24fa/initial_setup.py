"""
Initial Setup: Teacher's lesson plan template creation
Task ID: calc_grs_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Reference Data ---
    # The teacher has a raw reference sheet with subjects, levels, and schedule notes
    # but no formatted weekly template yet
    ws = wb.active
    ws.title = "Planning Notes"

    # Some raw planning notes the teacher jotted down
    ws["A1"] = "Subject List"
    ws["A2"] = "Math"
    ws["A3"] = "English"
    ws["A4"] = "Science"
    ws["A5"] = "History"
    ws["A6"] = "PE"

    ws["C1"] = "Class Levels"
    ws["C2"] = "7th Grade"
    ws["C3"] = "8th Grade"
    ws["C4"] = "6th Grade"

    ws["E1"] = "Weekly Schedule Notes"
    ws["E2"] = "School day: 8:00 AM - 3:00 PM"
    ws["E3"] = "Each period: 30 minutes"
    ws["E4"] = "Double periods for lab sessions (Science, PE)"
    ws["E5"] = "Monday staff meeting at 3:00 PM"
    ws["E6"] = "Need to track curriculum standards for district accreditation"
    ws["E7"] = "Color code subjects for quick visual reference"

    ws["A8"] = ""
    ws["A9"] = "Curriculum Standards to Cover This Week"
    ws["A10"] = "CCSS.MATH.6.RP.A.1 - Understand ratio concepts"
    ws["A11"] = "CCSS.MATH.7.EE.B.3 - Solve multi-step real-world problems"
    ws["A12"] = "CCSS.ELA.7.RL.1 - Cite textual evidence"
    ws["A13"] = "CCSS.ELA.8.W.3 - Write narratives"
    ws["A14"] = "NGSS.MS-PS1-1 - Develop models for atomic composition"
    ws["A15"] = "NGSS.MS-LS1-3 - Use argument for body as system of cells"
    ws["A16"] = "NCSS.D2.His.1.6-8 - Analyze connections among events"
    ws["A17"] = "SHAPE.S1.M8 - Apply movement concepts in modified games"

    # Basic header styling only
    for cell_ref in ["A1", "C1", "E1", "A9"]:
        ws[cell_ref].font = Font(bold=True)

    # Set reasonable column widths for readability
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["E"].width = 50

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
