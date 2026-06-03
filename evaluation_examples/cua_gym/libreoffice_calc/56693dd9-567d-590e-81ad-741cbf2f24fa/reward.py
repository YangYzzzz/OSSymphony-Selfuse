"""
Reward Script: Teacher's weekly lesson plan template in LibreOffice Calc
Task ID: calc_grs_040
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Two sheets exist: "Weekly Schedule" and "Curriculum Standards"
  Component 2 (0.25) - Schedule grid with time slots (30-min blocks 8am-3pm) in rows, Mon-Fri in cols
  Component 3 (0.20) - Color-coded subject cells (Math=light blue, English=light green, etc.)
  Component 4 (0.15) - Merged cells for double periods
  Component 5 (0.10) - Preparation notes section below schedule grid
  Component 6 (0.10) - Curriculum Standards sheet with X marks and lesson references
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_040'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Component 1: Two sheets - "Weekly Schedule" (or similar) and "Curriculum Standards" (or similar) (0.20 pts)
    # The initial file has only "Planning Notes" — so having these two sheets is a task-introduced change.
    try:
        has_schedule_sheet = False
        has_standards_sheet = False
        schedule_ws = None
        standards_ws = None

        for name in sheet_names:
            nl = name.lower()
            if 'schedule' in nl or 'weekly' in nl or 'lesson plan' in nl:
                has_schedule_sheet = True
                schedule_ws = wb[name]
            if 'standard' in nl or 'curriculum' in nl:
                has_standards_sheet = True
                standards_ws = wb[name]

        if has_schedule_sheet and has_standards_sheet:
            print(f"PASS: Component 1 - Both schedule and standards sheets found (0.20 pts)")
            total_score += 0.20
        elif has_schedule_sheet or has_standards_sheet:
            print(f"PARTIAL: Component 1 - Only one of the two required sheets found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 - Neither schedule nor standards sheet found. Sheets: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    if schedule_ws is None:
        print("CRITICAL: No schedule sheet found, cannot verify remaining components")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Schedule grid with time slots in rows and weekdays in columns (0.25 pts)
    # Initial file has no such grid. Check for: time labels in col A, day headers (Mon-Fri) in row header.
    try:
        comp2_score = 0.0

        # Check for day-of-week headers in the header row area (rows 1-3)
        weekdays = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
        found_days = set()
        for row_idx in range(1, 4):
            for col_idx in range(1, 10):
                val = schedule_ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str):
                    for day in weekdays:
                        if day in val.lower():
                            found_days.add(day)

        if len(found_days) >= 4:
            comp2_score += 0.10
            print(f"  PASS: Found {len(found_days)}/5 weekday headers")
        else:
            print(f"  FAIL: Only found {len(found_days)} weekday headers: {found_days}")

        # Check for time slots in column A (look for AM/PM patterns)
        time_count = 0
        for row_idx in range(1, schedule_ws.max_row + 1):
            val = schedule_ws.cell(row=row_idx, column=1).value
            if val and isinstance(val, str) and ('AM' in val.upper() or 'PM' in val.upper()):
                time_count += 1

        if time_count >= 8:  # At least 8 time slots for 30-min blocks 8am-3pm
            comp2_score += 0.10
            print(f"  PASS: Found {time_count} time slots in column A")
        elif time_count >= 4:
            comp2_score += 0.05
            print(f"  PARTIAL: Found {time_count} time slots (expected >= 8)")
        else:
            print(f"  FAIL: Found only {time_count} time slots in column A")

        # Check that schedule cells contain subject+grade+objective content (multiline)
        content_cells_with_subject = 0
        subjects = ['math', 'english', 'science', 'history', 'pe']
        for row_idx in range(2, min(schedule_ws.max_row + 1, 20)):
            for col_idx in range(2, 7):
                val = schedule_ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str):
                    val_lower = val.lower()
                    if any(s in val_lower for s in subjects) and 'grade' in val_lower:
                        content_cells_with_subject += 1

        if content_cells_with_subject >= 15:
            comp2_score += 0.05
            print(f"  PASS: Found {content_cells_with_subject} lesson cells with subject+grade")
        elif content_cells_with_subject >= 5:
            comp2_score += 0.025
            print(f"  PARTIAL: Found {content_cells_with_subject} lesson cells (expected >= 15)")
        else:
            print(f"  FAIL: Only {content_cells_with_subject} lesson cells with subject+grade content")

        if comp2_score > 0:
            print(f"PASS: Component 2 - Schedule grid verified ({comp2_score} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 - Schedule grid not properly structured")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Color-coded subject cells (0.20 pts)
    # Initial file has no color-coded cells. Check that cells with different subjects have different fill colors.
    try:
        subject_colors = {}  # subject -> set of fill colors seen
        for row_idx in range(2, min(schedule_ws.max_row + 1, 18)):
            for col_idx in range(2, 7):
                cell = schedule_ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val and isinstance(val, str) and not isinstance(cell, MergedCell):
                    val_lower = val.lower()
                    fill_rgb = None
                    try:
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                            rgb = cell.fill.fgColor.rgb
                            if rgb != '00000000':
                                fill_rgb = rgb
                    except:
                        pass

                    if fill_rgb:
                        for subj in ['math', 'english', 'science', 'history', 'pe']:
                            if val_lower.startswith(subj):
                                if subj not in subject_colors:
                                    subject_colors[subj] = set()
                                subject_colors[subj].add(fill_rgb)

        # Verify: at least 4 subjects have color-coding, and different subjects have different colors
        subjects_with_color = len(subject_colors)
        # Check that different subjects have distinct colors
        all_colors = set()
        for subj, colors in subject_colors.items():
            all_colors.update(colors)

        distinct_colors = len(all_colors)

        if subjects_with_color >= 4 and distinct_colors >= 4:
            print(f"PASS: Component 3 - {subjects_with_color} subjects color-coded with {distinct_colors} distinct colors (0.20 pts)")
            total_score += 0.20
        elif subjects_with_color >= 3 and distinct_colors >= 3:
            print(f"PARTIAL: Component 3 - {subjects_with_color} subjects, {distinct_colors} colors (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 - Only {subjects_with_color} subjects with color, {distinct_colors} distinct colors")
            print(f"  Subject colors found: {subject_colors}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Merged cells for double periods (0.15 pts)
    # Initial file has no merged cells. Check that the schedule sheet has merged cell ranges.
    try:
        merged_ranges = list(schedule_ws.merged_cells.ranges)
        # Filter for merged ranges in the schedule area (rows 2-17, cols B-F)
        schedule_merges = []
        for mr in merged_ranges:
            if mr.min_row >= 2 and mr.max_row <= 17 and mr.min_col >= 2 and mr.max_col <= 6:
                # Must span at least 2 rows (double period = consecutive time slots merged)
                if mr.max_row - mr.min_row >= 1:
                    schedule_merges.append(str(mr))

        if len(schedule_merges) >= 3:
            print(f"PASS: Component 4 - Found {len(schedule_merges)} double-period merges in schedule area (0.15 pts)")
            total_score += 0.15
        elif len(schedule_merges) >= 1:
            print(f"PARTIAL: Component 4 - Found {len(schedule_merges)} merges (expected >= 3) (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 - No double-period merged cells found in schedule grid")
            print(f"  All merged ranges: {[str(m) for m in merged_ranges]}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Preparation notes section below schedule grid (0.10 pts)
    # Initial file has no prep notes section. Look for "prep" or "notes" label below the schedule area.
    try:
        found_prep_section = False
        prep_row = None
        for row_idx in range(15, schedule_ws.max_row + 1):
            val = schedule_ws.cell(row=row_idx, column=1).value
            if val and isinstance(val, str):
                vl = val.lower()
                if 'prep' in vl or 'note' in vl:
                    found_prep_section = True
                    prep_row = row_idx
                    break

        if found_prep_section:
            # Check that there's actual content in the prep notes section
            notes_content_count = 0
            for row_idx in range(prep_row + 1, min(prep_row + 10, schedule_ws.max_row + 1)):
                for col_idx in range(1, 7):
                    val = schedule_ws.cell(row=row_idx, column=col_idx).value
                    if val and isinstance(val, str) and len(val.strip()) > 10:
                        notes_content_count += 1
                        break  # count one per row

            if notes_content_count >= 3:
                print(f"PASS: Component 5 - Prep notes section at row {prep_row} with {notes_content_count} note entries (0.10 pts)")
                total_score += 0.10
            elif notes_content_count >= 1:
                print(f"PARTIAL: Component 5 - Prep notes section found but only {notes_content_count} entries (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 - Prep notes header found but no content")
        else:
            print(f"FAIL: Component 5 - No preparation notes section found below schedule")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Curriculum Standards sheet with X marks and lesson references (0.10 pts)
    # Initial file has no such sheet. Verify structure and X marks.
    try:
        if standards_ws is None:
            print(f"FAIL: Component 6 - No curriculum standards sheet found")
        else:
            # Check for standard codes, X marks, and references
            x_marks = 0
            standard_codes = 0
            lesson_refs = 0
            for row_idx in range(2, standards_ws.max_row + 1):
                # Standard code in col A
                code_val = standards_ws.cell(row=row_idx, column=1).value
                if code_val and isinstance(code_val, str) and len(code_val.strip()) > 3:
                    standard_codes += 1

                # X mark in col D (or wherever "Covered" column is)
                for col_idx in range(1, standards_ws.max_column + 1):
                    val = standards_ws.cell(row=row_idx, column=col_idx).value
                    if val and str(val).strip().upper() == 'X':
                        x_marks += 1
                        break

                # Lesson reference (look for time references like "Mon", "Tue", etc.)
                for col_idx in range(1, standards_ws.max_column + 1):
                    val = standards_ws.cell(row=row_idx, column=col_idx).value
                    if val and isinstance(val, str) and ('Mon' in val or 'Tue' in val or 'AM' in val or 'PM' in val):
                        lesson_refs += 1
                        break

            if standard_codes >= 5 and x_marks >= 5:
                if lesson_refs >= 3:
                    print(f"PASS: Component 6 - {standard_codes} standards, {x_marks} X marks, {lesson_refs} lesson refs (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"PARTIAL: Component 6 - Standards and X marks found but only {lesson_refs} lesson refs (0.07 pts)")
                    total_score += 0.07
            elif standard_codes >= 3 and x_marks >= 3:
                print(f"PARTIAL: Component 6 - Found {standard_codes} standards, {x_marks} X marks (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 6 - Insufficient standards ({standard_codes}) or X marks ({x_marks})")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(total_score, 1.0)
    final_score = round(final_score, 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
