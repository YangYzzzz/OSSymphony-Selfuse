"""
Initial Setup: Paste Special formatting from header row to other sections
Task ID: calc_gsi_027
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Data"

    # --- Define formatting for Section 1 header only ---
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Column headers used across all sections
    headers = ["Employee", "Department", "Revenue ($)", "Target ($)", "Rating"]

    # Column widths for readability
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12

    # ========== Section 1: Q1 Performance (rows 1-8) ==========
    # Header row 1 — FORMATTED (this is the source for Paste Special)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    q1_data = [
        ["Sarah Chen", "Engineering", 87500, 80000, "Excellent"],
        ["Marcus Johnson", "Marketing", 62300, 65000, "Good"],
        ["Elena Rodriguez", "Sales", 94200, 90000, "Excellent"],
        ["James Park", "Engineering", 78100, 80000, "Good"],
        ["Aisha Patel", "Finance", 55800, 58000, "Satisfactory"],
        ["David Kim", "Sales", 101400, 95000, "Excellent"],
        ["Rachel Foster", "Marketing", 67900, 65000, "Good"],
    ]
    for r, row_data in enumerate(q1_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # ========== Section 2: Q2 Performance (rows 10-17) ==========
    # Header row 10 — UNFORMATTED (destination for paste special)
    for col, h in enumerate(headers, 1):
        ws.cell(row=10, column=col, value=h)

    q2_data = [
        ["Sarah Chen", "Engineering", 91200, 85000, "Excellent"],
        ["Marcus Johnson", "Marketing", 68400, 68000, "Good"],
        ["Elena Rodriguez", "Sales", 99800, 95000, "Excellent"],
        ["James Park", "Engineering", 82500, 82000, "Good"],
        ["Aisha Patel", "Finance", 59300, 60000, "Satisfactory"],
        ["David Kim", "Sales", 108700, 100000, "Excellent"],
        ["Rachel Foster", "Marketing", 71200, 70000, "Good"],
    ]
    for r, row_data in enumerate(q2_data, 11):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # ========== Section 3: Q3 Performance (rows 19-26) ==========
    # Header row 19 — UNFORMATTED (destination for paste special)
    for col, h in enumerate(headers, 1):
        ws.cell(row=19, column=col, value=h)

    q3_data = [
        ["Sarah Chen", "Engineering", 95400, 90000, "Excellent"],
        ["Marcus Johnson", "Marketing", 72100, 72000, "Good"],
        ["Elena Rodriguez", "Sales", 103500, 100000, "Excellent"],
        ["James Park", "Engineering", 86800, 85000, "Good"],
        ["Aisha Patel", "Finance", 62700, 62000, "Good"],
        ["David Kim", "Sales", 112300, 105000, "Excellent"],
        ["Rachel Foster", "Marketing", 74800, 73000, "Good"],
    ]
    for r, row_data in enumerate(q3_data, 20):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # ========== Section 4: Q4 Performance (rows 28-35) ==========
    # Header row 28 — UNFORMATTED (destination for paste special)
    for col, h in enumerate(headers, 1):
        ws.cell(row=28, column=col, value=h)

    q4_data = [
        ["Sarah Chen", "Engineering", 98700, 95000, "Excellent"],
        ["Marcus Johnson", "Marketing", 75600, 75000, "Good"],
        ["Elena Rodriguez", "Sales", 107200, 105000, "Excellent"],
        ["James Park", "Engineering", 90100, 88000, "Good"],
        ["Aisha Patel", "Finance", 65400, 65000, "Good"],
        ["David Kim", "Sales", 115800, 110000, "Excellent"],
        ["Rachel Foster", "Marketing", 78300, 76000, "Good"],
    ]
    for r, row_data in enumerate(q4_data, 29):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
