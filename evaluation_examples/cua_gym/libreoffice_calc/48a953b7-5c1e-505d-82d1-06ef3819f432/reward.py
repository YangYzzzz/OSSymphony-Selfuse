"""
Reward Script: Paste Special - Formatting Only
Task ID: calc_gsi_027
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): Row 10 header cells (A10:E10) have source formatting applied
  - Component 2 (0.35): Row 19 header cells (A19:E19) have source formatting applied
  - Component 3 (0.30): Row 28 header cells (A28:E28) have source formatting applied

The task requires using Paste Special to paste ONLY formatting from the header row (row 1)
to three other section headers (rows 10, 19, 28). The text content must remain intact
while the visual formatting (background color, font bold/name, font color, borders,
alignment) is applied.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_027'

# Expected formatting properties from the source header row (row 1)
EXPECTED_FILL_FG = 'FF4472C4'  # blue background (ARGB)
EXPECTED_FONT_BOLD = True
EXPECTED_FONT_NAME = 'Arial'
EXPECTED_BORDER_STYLE = 'thin'  # all four sides
EXPECTED_ALIGNMENT_H = 'center'
EXPECTED_ALIGNMENT_V = 'center'

# The three target header rows that should receive formatting
TARGET_HEADER_ROWS = [10, 19, 28]
COLUMNS = ['A', 'B', 'C', 'D', 'E']

# Expected header text values (must remain unchanged after paste-formatting)
EXPECTED_HEADERS = ['Employee', 'Department', 'Revenue ($)', 'Target ($)', 'Rating']


def check_cell_formatting(ws, coord):
    """
    Check if a cell has the expected header formatting.
    Returns a tuple (formatting_score, details_str).
    Checks: fill color, bold, font name, borders, alignment.
    Each sub-check is worth 1/5 of the cell's contribution.
    """
    cell = ws[coord]
    checks_passed = 0
    total_checks = 5
    details = []

    # 1. Background fill color
    try:
        fg_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        if fg_rgb == EXPECTED_FILL_FG:
            checks_passed += 1
        else:
            details.append(f"fill={fg_rgb} (expected {EXPECTED_FILL_FG})")
    except Exception:
        details.append("fill=error")

    # 2. Font bold
    if cell.font.bold == EXPECTED_FONT_BOLD:
        checks_passed += 1
    else:
        details.append(f"bold={cell.font.bold}")

    # 3. Font name
    if cell.font.name == EXPECTED_FONT_NAME:
        checks_passed += 1
    else:
        details.append(f"font_name={cell.font.name} (expected {EXPECTED_FONT_NAME})")

    # 4. Borders (all four sides should be thin)
    border_ok = (
        cell.border.top.style == EXPECTED_BORDER_STYLE
        and cell.border.bottom.style == EXPECTED_BORDER_STYLE
        and cell.border.left.style == EXPECTED_BORDER_STYLE
        and cell.border.right.style == EXPECTED_BORDER_STYLE
    )
    if border_ok:
        checks_passed += 1
    else:
        details.append(f"borders=({cell.border.top.style},{cell.border.bottom.style},"
                       f"{cell.border.left.style},{cell.border.right.style})")

    # 5. Alignment (horizontal=center, vertical=center)
    align_ok = (
        cell.alignment.horizontal == EXPECTED_ALIGNMENT_H
        and cell.alignment.vertical == EXPECTED_ALIGNMENT_V
    )
    if align_ok:
        checks_passed += 1
    else:
        details.append(f"align=({cell.alignment.horizontal},{cell.alignment.vertical})")

    return checks_passed / total_checks, details


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition gate: verify the file has the expected sheet and dimensions
    if ws.max_row < 35 or ws.max_column < 5:
        print(f"FAIL: Unexpected dimensions: {ws.max_row} rows x {ws.max_column} cols (expected >= 35x5)")
        print("REWARD: 0.0")
        return 0.0

    # Weight per target header row
    weights = {10: 0.35, 19: 0.35, 28: 0.30}

    for row_num in TARGET_HEADER_ROWS:
        weight = weights[row_num]
        component_label = f"Row {row_num} header formatting"

        try:
            row_score = 0.0
            row_fails = []
            cells_checked = 0

            for i, col in enumerate(COLUMNS):
                coord = f"{col}{row_num}"
                cell = ws[coord]

                # First verify that text content is preserved (precondition, not scored)
                expected_text = EXPECTED_HEADERS[i]
                if cell.value != expected_text:
                    print(f"WARN: {coord} text changed from '{expected_text}' to '{cell.value}' — "
                          f"content should be preserved; skipping formatting check for this cell")
                    continue

                cells_checked += 1
                cell_fmt_score, cell_details = check_cell_formatting(ws, coord)
                row_score += cell_fmt_score

                if cell_details:
                    row_fails.append(f"{coord}: {', '.join(cell_details)}")

            if cells_checked > 0:
                # Normalize: average formatting score across cells checked
                avg_score = row_score / cells_checked
                component_pts = weight * avg_score
                total_score += component_pts

                if avg_score >= 1.0:
                    print(f"PASS: {component_label} — all {cells_checked} cells fully formatted ({component_pts:.3f} pts)")
                else:
                    print(f"PARTIAL: {component_label} — avg format score {avg_score:.2f} ({component_pts:.3f} pts)")
                    for fail in row_fails:
                        print(f"  - {fail}")
            else:
                print(f"FAIL: {component_label} — no cells with expected text found")

        except Exception as e:
            print(f"ERROR: {component_label} — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved GUI changes before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
