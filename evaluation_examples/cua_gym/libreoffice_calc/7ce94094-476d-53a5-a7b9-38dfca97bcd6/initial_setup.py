"""
Initial Setup: Build a simple linear regression forecast using SLOPE and INTERCEPT functions
Task ID: calc_ops_090
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Regression ---
    ws = wb.active
    ws.title = 'Regression'

    # Column headers
    ws.cell(row=1, column=1, value='Quarter')
    ws.cell(row=1, column=2, value='Period')
    ws.cell(row=1, column=3, value='Demand')

    # Header styling
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    white_font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows
    data = [
        ['Q1 2025', 1, 1000],
        ['Q2 2025', 2, 1150],
        ['Q3 2025', 3, 1280],
        ['Q4 2025', 4, 1400],
        ['Q1 2026', 5, 1520],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Center-align period column
    for r in range(2, 7):
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="center")

    # Number format for demand column
    for r in range(2, 7):
        ws.cell(row=r, column=3).number_format = '#,##0'

    # Analysis labels in column E
    ws.cell(row=1, column=5, value='Analysis')
    ws['E1'].font = Font(bold=True, size=12, name='Calibri')

    ws.cell(row=2, column=5, value='Slope')
    ws.cell(row=3, column=5, value='Intercept')
    ws.cell(row=4, column=5, value='R-squared')
    ws.cell(row=5, column=5, value='Q2 2026 Forecast (Period 6)')

    # Style analysis labels
    label_font = Font(bold=True, size=11, name='Calibri')
    for r in range(2, 6):
        ws.cell(row=r, column=5).font = label_font

    # F2:F5 intentionally left EMPTY - the task is to fill these with formulas

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 4
    ws.column_dimensions['E'].width = 32
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
