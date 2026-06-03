"""
Reward Script: Linear regression forecast using SLOPE and INTERCEPT functions
Task ID: calc_ops_090
Domain: libreoffice_calc
Scoring:
  Component 1 - SLOPE formula in F2 (0.25 pts)
  Component 2 - INTERCEPT formula in F3 (0.25 pts)
  Component 3 - RSQ formula in F4 (0.25 pts)
  Component 4 - Forecast formula in F5 combining SLOPE*6+INTERCEPT (0.25 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_090'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Regression' sheet must exist
    if 'Regression' not in wb.sheetnames:
        print("FAIL: 'Regression' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Regression']

    # Component 1: SLOPE formula in F2 (0.25 points)
    # Expected: =SLOPE(C2:C6,B2:B6)
    try:
        val_f2 = ws['F2'].value
        norm_f2 = normalize_formula(val_f2)
        if norm_f2 == normalize_formula('=SLOPE(C2:C6,B2:B6)'):
            print(f"PASS: Component 1 — SLOPE formula in F2: {val_f2} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected =SLOPE(C2:C6,B2:B6) in F2, found: {val_f2}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: INTERCEPT formula in F3 (0.25 points)
    # Expected: =INTERCEPT(C2:C6,B2:B6)
    try:
        val_f3 = ws['F3'].value
        norm_f3 = normalize_formula(val_f3)
        if norm_f3 == normalize_formula('=INTERCEPT(C2:C6,B2:B6)'):
            print(f"PASS: Component 2 — INTERCEPT formula in F3: {val_f3} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Expected =INTERCEPT(C2:C6,B2:B6) in F3, found: {val_f3}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: RSQ formula in F4 (0.25 points)
    # Expected: =RSQ(C2:C6,B2:B6)
    try:
        val_f4 = ws['F4'].value
        norm_f4 = normalize_formula(val_f4)
        if norm_f4 == normalize_formula('=RSQ(C2:C6,B2:B6)'):
            print(f"PASS: Component 3 — RSQ formula in F4: {val_f4} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Expected =RSQ(C2:C6,B2:B6) in F4, found: {val_f4}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Forecast formula in F5 (0.25 points)
    # Expected: =SLOPE(C2:C6,B2:B6)*6+INTERCEPT(C2:C6,B2:B6)
    # The agent might also write it as =F2*6+F3 or similar valid variants
    try:
        val_f5 = ws['F5'].value
        norm_f5 = normalize_formula(val_f5)
        # Accept the full inline formula
        expected_inline = normalize_formula('=SLOPE(C2:C6,B2:B6)*6+INTERCEPT(C2:C6,B2:B6)')
        # Also accept cell-reference variant: =F2*6+F3
        expected_ref = normalize_formula('=F2*6+F3')
        if norm_f5 in (expected_inline, expected_ref):
            print(f"PASS: Component 4 — Forecast formula in F5: {val_f5} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — Expected forecast formula in F5, found: {val_f5}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
