"""
Initial Setup: Dynamic pricing system using INDEX/MATCH nested lookup
Task ID: calc_gen_lookup_061
Domain: libreoffice_calc

Creates:
- Sheet 'Orders': 200 orders with Order ID, Product ID, Customer Tier, Qty (Unit Price and Total left EMPTY)
- Sheet 'PriceTable': 50 products with Standard/Gold/Silver/Bronze pricing (some tier columns blank)
"""

import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_lookup_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

# Product IDs and names
PRODUCTS = [
    ('PRD-001', 'Laptop Pro 15'),
    ('PRD-002', 'Wireless Mouse'),
    ('PRD-003', 'Mechanical Keyboard'),
    ('PRD-004', 'USB-C Hub'),
    ('PRD-005', '4K Monitor'),
    ('PRD-006', 'Noise Cancelling Headphones'),
    ('PRD-007', 'Webcam HD'),
    ('PRD-008', 'External SSD 1TB'),
    ('PRD-009', 'Docking Station'),
    ('PRD-010', 'Standing Desk'),
    ('PRD-011', 'Ergonomic Chair'),
    ('PRD-012', 'LED Desk Lamp'),
    ('PRD-013', 'Surge Protector'),
    ('PRD-014', 'Dual Monitor Arm'),
    ('PRD-015', 'Wireless Charger'),
    ('PRD-016', 'Smart Speaker'),
    ('PRD-017', 'Portable Projector'),
    ('PRD-018', 'Drawing Tablet'),
    ('PRD-019', 'Video Capture Card'),
    ('PRD-020', 'Microphone Studio'),
    ('PRD-021', 'Cable Management Kit'),
    ('PRD-022', 'Anti-Glare Screen Filter'),
    ('PRD-023', 'Laptop Stand'),
    ('PRD-024', 'USB-A Hub 7-Port'),
    ('PRD-025', 'Trackball Mouse'),
    ('PRD-026', 'Compact Keyboard'),
    ('PRD-027', '27-inch Monitor'),
    ('PRD-028', 'Mechanical Switch Tester'),
    ('PRD-029', 'Monitor Calibrator'),
    ('PRD-030', 'HDMI 2.1 Cable'),
    ('PRD-031', 'Fiber Optic Cable 10m'),
    ('PRD-032', 'KVM Switch 2-Port'),
    ('PRD-033', 'Desk Mat XL'),
    ('PRD-034', 'Phone Mount'),
    ('PRD-035', 'Wrist Rest Pad'),
    ('PRD-036', 'Network Switch 8-Port'),
    ('PRD-037', 'Powerline Adapter Kit'),
    ('PRD-038', 'WiFi 6 Router'),
    ('PRD-039', 'NAS Storage 4-Bay'),
    ('PRD-040', 'UPS 1500VA'),
    ('PRD-041', 'Label Printer'),
    ('PRD-042', 'Barcode Scanner'),
    ('PRD-043', 'Receipt Printer'),
    ('PRD-044', 'Document Scanner'),
    ('PRD-045', 'All-in-One Printer'),
    ('PRD-046', 'Shredder 12-Sheet'),
    ('PRD-047', 'Laminator A4'),
    ('PRD-048', 'Binding Machine'),
    ('PRD-049', 'Electric Stapler'),
    ('PRD-050', 'Desk Organizer Pro'),
]

TIERS = ['Gold', 'Silver', 'Bronze', 'Standard']

FIRST_NAMES = [
    'Aiden', 'Sofia', 'Marcus', 'Elena', 'James', 'Priya', 'Noah', 'Yuki',
    'Liam', 'Zara', 'Ethan', 'Mei', 'Oliver', 'Fatima', 'Lucas', 'Nadia',
    'Mason', 'Ingrid', 'Logan', 'Amara'
]
LAST_NAMES = [
    'Thornton', 'Vasquez', 'Okafor', 'Petrov', 'Hartmann', 'Nakamura',
    'Chen', 'Williams', 'Patel', 'Rivera', 'Murphy', 'Andersen', 'Kowalski',
    'Santos', 'Nguyen', 'Fischer', 'Martin', 'Diaz', 'Kim', 'Johnson'
]

COMPANIES = [
    'Apex Solutions', 'BlueStar Corp', 'Crestview Tech', 'Delta Systems',
    'Emerald Group', 'Fusion Dynamics', 'Global Reach Ltd', 'Horizon Partners',
    'Innovate Inc', 'Jetstream Co', 'Keystone Ventures', 'Luminary Labs',
    'Momentum Corp', 'Nexus Digital', 'Omega Networks', 'Pinnacle Group',
    'Quantum Works', 'Radiant Tech', 'Summit Enterprises', 'Titan Industries'
]


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Orders ----
    ws_orders = wb.active
    ws_orders.title = 'Orders'

    headers_orders = ['Order ID', 'Product ID', 'Customer Tier', 'Qty', 'Unit Price', 'Total']
    for col, h in enumerate(headers_orders, 1):
        ws_orders.cell(row=1, column=col, value=h)

    # Generate 200 realistic orders
    # Columns A-D populated; E (Unit Price) and F (Total) left EMPTY
    for i in range(2, 202):
        order_num = 10000 + i - 2
        order_id = f'ORD-{order_num}'
        product = random.choice(PRODUCTS)
        product_id = product[0]
        tier = random.choice(TIERS)
        qty = random.randint(1, 50)

        ws_orders.cell(row=i, column=1, value=order_id)     # A: Order ID
        ws_orders.cell(row=i, column=2, value=product_id)   # B: Product ID
        ws_orders.cell(row=i, column=3, value=tier)         # C: Customer Tier
        ws_orders.cell(row=i, column=4, value=qty)          # D: Qty
        # E (Unit Price) and F (Total) intentionally left EMPTY

    # ---- Sheet 2: PriceTable ----
    ws_price = wb.create_sheet('PriceTable')

    headers_price = ['Product ID', 'Standard', 'Gold', 'Silver', 'Bronze']
    for col, h in enumerate(headers_price, 1):
        ws_price.cell(row=1, column=col, value=h)

    # Generate 50 products with tier-based pricing
    # Some tier columns are blank (to test fallback to Standard)
    for i, (prod_id, prod_name) in enumerate(PRODUCTS, 2):
        standard_price = round(random.uniform(15.0, 1200.0), 2)
        # Gold: ~15% discount from Standard (always present for first 40; missing for some)
        # Silver: ~8% discount; Bronze: ~3% discount — some deliberately left blank
        gold_price = round(standard_price * 0.85, 2) if random.random() > 0.15 else None
        silver_price = round(standard_price * 0.92, 2) if random.random() > 0.20 else None
        bronze_price = round(standard_price * 0.97, 2) if random.random() > 0.25 else None

        ws_price.cell(row=i, column=1, value=prod_id)
        ws_price.cell(row=i, column=2, value=standard_price)
        ws_price.cell(row=i, column=3, value=gold_price)     # May be None (blank)
        ws_price.cell(row=i, column=4, value=silver_price)   # May be None (blank)
        ws_price.cell(row=i, column=5, value=bronze_price)   # May be None (blank)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Orders sheet: 200 rows (E and F columns empty)')
    print(f'  PriceTable sheet: 50 products with tier pricing (some tier prices blank for fallback testing)')


create_initial()
