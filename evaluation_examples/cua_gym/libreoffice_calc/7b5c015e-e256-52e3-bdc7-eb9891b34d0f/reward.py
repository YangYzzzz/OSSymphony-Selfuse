"""
Reward Script: Build a dynamic pricing system using INDEX/MATCH with tier-based lookup
Task ID: calc_gen_lookup_061
Domain: libreoffice_calc
Scoring:
  Component 1: E2:E201 have IFERROR-wrapped INDEX/MATCH formulas for price lookup (0.4 pts)
  Component 2: F2:F201 have Qty * Unit Price formulas (0.3 pts)
  Component 3: E and F columns formatted as currency (0.3 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_lookup_061'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task requires:
    - E2:E201 (Unit Price): IFERROR-wrapped price lookup using INDEX/MATCH
      matching Product ID in PriceTable and selecting column by Customer Tier;
      fallback to Standard price if tier price is blank; 'Price Not Found' on error
    - F2:F201 (Total): = D{row} * E{row}
    - E and F columns formatted as currency
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: required sheets must exist
    if 'Orders' not in wb.sheetnames or 'PriceTable' not in wb.sheetnames:
        print(f"FAIL: Required sheets missing. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_orders = wb['Orders']

    # -----------------------------------------------------------------------
    # Component 1: E2:E201 contain IFERROR-wrapped INDEX/MATCH price lookup (0.4 pts)
    # The formula must:
    #   - Start with =IFERROR(
    #   - Use INDEX and MATCH (or LET/CHOOSE as intermediate) for lookup
    #   - Reference PriceTable
    #   - Reference both Product ID (column B) and Customer Tier (column C)
    # In the initial file, E2:E201 are all None, so this FAILS on initial (correct).
    # -----------------------------------------------------------------------
    try:
        e_formula_count = 0
        e_iferror_count = 0
        e_index_match_count = 0
        e_pricetable_count = 0
        e_tier_count = 0
        e_fallback_count = 0

        for row in range(2, 202):
            cell_val = ws_orders.cell(row=row, column=5).value
            if cell_val is None:
                continue
            cell_str = str(cell_val).upper().replace(' ', '')
            e_formula_count += 1

            # Check IFERROR wrapping
            if cell_str.startswith('=IFERROR('):
                e_iferror_count += 1

            # Check INDEX and MATCH usage (handles LET/CHOOSE variants too)
            if 'INDEX(' in cell_str and 'MATCH(' in cell_str:
                e_index_match_count += 1

            # Check PriceTable reference
            if 'PRICETABLE' in cell_str:
                e_pricetable_count += 1

            # Check tier column reference (C column = Customer Tier)
            if ',C' in cell_str or '(C' in cell_str:
                e_tier_count += 1

            # Check fallback to standard (either IF(tier_price="", or CHOOSE/column 2 reference)
            if 'STANDARD' in cell_str or ',2)' in cell_str or ',2,' in cell_str or 'IF(' in cell_str:
                e_fallback_count += 1

        print(f"E formula diagnostics: total={e_formula_count}, iferror={e_iferror_count}, "
              f"index_match={e_index_match_count}, pricetable={e_pricetable_count}, "
              f"tier_ref={e_tier_count}, fallback={e_fallback_count}")

        # All 200 rows must have formulas with IFERROR + INDEX/MATCH + PriceTable ref
        all_correct = (e_formula_count == 200 and e_iferror_count == 200
                       and e_index_match_count == 200 and e_pricetable_count == 200)
        if all_correct:
            print(f"PASS: Component 1 — All 200 Unit Price formulas use IFERROR-wrapped INDEX/MATCH (0.4 pts)")
            total_score += 0.4
        elif e_formula_count == 200 and e_iferror_count == 200:
            # Has IFERROR but may not use INDEX/MATCH (e.g., VLOOKUP variant) — partial
            print(f"PARTIAL: Component 1 — All 200 rows have IFERROR formulas but INDEX/MATCH not confirmed ({e_index_match_count}/200 rows). (0.2 pts)")
            total_score += 0.2
        elif e_formula_count > 0:
            print(f"FAIL: Component 1 — Only {e_formula_count}/200 rows have Unit Price formulas "
                  f"(IFERROR: {e_iferror_count}, INDEX/MATCH: {e_index_match_count})")
        else:
            print(f"FAIL: Component 1 — No Unit Price formulas found in E2:E201")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: F2:F201 contain Qty * Unit Price totals (0.3 pts)
    # Each cell in F column should be =D{row}*E{row}
    # In the initial file, F2:F201 are all None, so this FAILS on initial (correct).
    # -----------------------------------------------------------------------
    try:
        f_formula_count = 0
        f_correct_count = 0
        f_multiply_count = 0

        for row in range(2, 202):
            cell_val = ws_orders.cell(row=row, column=6).value
            if cell_val is None:
                continue
            cell_str = str(cell_val).upper().replace(' ', '')
            f_formula_count += 1

            # Exact pattern: =D{row}*E{row} (in any order)
            expected1 = f'=D{row}*E{row}'
            expected2 = f'=E{row}*D{row}'
            if (str(cell_val).upper().replace(' ', '') == expected1.upper()
                    or str(cell_val).upper().replace(' ', '') == expected2.upper()):
                f_correct_count += 1
            elif '*' in cell_str and f'D{row}' in cell_str and f'E{row}' in cell_str:
                f_multiply_count += 1

        print(f"F formula diagnostics: total={f_formula_count}, correct_D*E={f_correct_count}, "
              f"other_multiply={f_multiply_count}")

        if f_correct_count == 200:
            print(f"PASS: Component 2 — All 200 Total formulas are =D{{row}}*E{{row}} (0.3 pts)")
            total_score += 0.3
        elif f_correct_count + f_multiply_count == 200:
            print(f"PASS: Component 2 — All 200 Total formulas multiply D*E (alternative form) (0.3 pts)")
            total_score += 0.3
        elif f_correct_count + f_multiply_count > 0:
            covered = f_correct_count + f_multiply_count
            print(f"PARTIAL: Component 2 — Only {covered}/200 Total formulas correct. (0.15 pts)")
            total_score += 0.15
        elif f_formula_count > 0:
            print(f"FAIL: Component 2 — {f_formula_count} Total cells have values but don't match D*E pattern")
        else:
            print(f"FAIL: Component 2 — No Total formulas found in F2:F201")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: E and F columns are formatted as currency (0.3 pts)
    # In the initial file, E and F are empty with no special format.
    # After task: both should have a currency-like number format (contains '$' or 'currency').
    # -----------------------------------------------------------------------
    try:
        e_currency_count = 0
        f_currency_count = 0
        currency_keywords = ['$', 'currency', '#,##0']

        for row in range(2, 202):
            e_fmt = ws_orders.cell(row=row, column=5).number_format
            f_fmt = ws_orders.cell(row=row, column=6).number_format
            # Check for currency-style format
            if any(kw in str(e_fmt) for kw in currency_keywords):
                e_currency_count += 1
            if any(kw in str(f_fmt) for kw in currency_keywords):
                f_currency_count += 1

        print(f"Currency format diagnostics: E currency rows={e_currency_count}/200, F currency rows={f_currency_count}/200")

        if e_currency_count == 200 and f_currency_count == 200:
            print(f"PASS: Component 3 — All E and F cells formatted as currency (0.3 pts)")
            total_score += 0.3
        elif e_currency_count >= 190 and f_currency_count >= 190:
            print(f"PASS: Component 3 — Nearly all E and F cells formatted as currency (0.3 pts)")
            total_score += 0.3
        elif e_currency_count > 0 or f_currency_count > 0:
            print(f"PARTIAL: Component 3 — Currency format partial: E={e_currency_count}/200, F={f_currency_count}/200 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — No currency formatting found in E or F columns")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
