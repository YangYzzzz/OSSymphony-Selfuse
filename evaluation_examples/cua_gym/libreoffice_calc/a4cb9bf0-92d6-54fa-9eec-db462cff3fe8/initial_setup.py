"""
Initial Setup: Monthly Expense Report - blank template with desktop reference files
Task ID: osworld_multi_apps_receipt_to_calc_011
Domain: libreoffice_calc

Creates:
  - /home/user/monthly_report.xlsx (blank template with 3 empty sheets)
  - /home/user/Desktop/credit_card_statement.pdf (credit card transactions)
  - /home/user/Desktop/petty_cash_photo.jpg (petty cash log image)
  - /home/user/Desktop/receipt_001.jpg through receipt_006.jpg (receipt images)
  - Opens monthly_report.xlsx in LibreOffice Calc
"""

import os
import shlex
import subprocess
import sys
import time

# Install required packages on the VM
subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl', 'fpdf2', 'Pillow', '--quiet'],
               capture_output=True)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── paths ──────────────────────────────────────────────────────────────────────
WORKDIR    = '/home/user'
DESKTOP    = '/home/user/Desktop'
TASK_ID    = 'osworld_multi_apps_receipt_to_calc_011'
OUTPUT     = f'{WORKDIR}/monthly_report.xlsx'

# ── helpers ────────────────────────────────────────────────────────────────────
def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch a GUI app on the VM display (non-blocking)."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# ── 1. Create blank monthly_report.xlsx ───────────────────────────────────────
def create_blank_workbook():
    wb = openpyxl.Workbook()

    # --- Sheet1: All Transactions (headers only, NO data) ---
    ws1 = wb.active
    ws1.title = 'All Transactions'
    headers1 = ['Date', 'Description', 'Category', 'Amount', 'Source', 'Tax Deductible']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 18
    ws1.column_dimensions['F'].width = 16
    ws1.freeze_panes = 'A2'

    # --- Sheet2: By Category (headers only, NO SUMIF, NO chart) ---
    ws2 = wb.create_sheet('By Category')
    headers2 = ['Category', 'Total Amount', '% of Total']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 14

    # --- Sheet3: Tax Deductible (headers only, NO data) ---
    ws3 = wb.create_sheet('Tax Deductible')
    headers3 = ['Date', 'Description', 'Category', 'Amount', 'Source']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 30
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 12
    ws3.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Blank workbook created: {OUTPUT}')


# ── 2. Create desktop reference files ─────────────────────────────────────────
def create_desktop_files():
    os.makedirs(DESKTOP, exist_ok=True)

    # --- credit_card_statement.pdf ---
    _create_pdf_credit_card()

    # --- petty_cash_photo.jpg ---
    _create_petty_cash_jpg()

    # --- receipt images ---
    _create_receipt_images()


def _create_pdf_credit_card():
    """Create a realistic credit card statement PDF."""
    try:
        from fpdf import FPDF
    except ImportError:
        subprocess.run(['pip3', 'install', 'fpdf2'], capture_output=True)
        from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 16)
    pdf.cell(0, 10, 'VISA Platinum - Monthly Statement', ln=True, align='C')
    pdf.set_font('Helvetica', '', 10)
    pdf.cell(0, 6, 'Statement Period: March 1-31, 2025     Account: ****-4821', ln=True, align='C')
    pdf.ln(4)

    pdf.set_font('Helvetica', 'B', 11)
    pdf.set_fill_color(220, 230, 245)
    pdf.cell(30, 8, 'Date', border=1, fill=True)
    pdf.cell(80, 8, 'Description', border=1, fill=True)
    pdf.cell(25, 8, 'Category', border=1, fill=True)
    pdf.cell(30, 8, 'Amount ($)', border=1, fill=True, ln=True)

    transactions = [
        ('2025-03-02', 'Amazon Web Services',         'Software',     234.50),
        ('2025-03-04', 'Delta Airlines Flight DL1204', 'Travel',       487.00),
        ('2025-03-05', 'Marriott Hotel Chicago',       'Accommodation',312.75),
        ('2025-03-07', 'Uber Business Trip',           'Transport',     38.20),
        ('2025-03-09', 'Office Depot Supplies',        'Office',        76.40),
        ('2025-03-11', 'Adobe Creative Cloud',         'Software',      54.99),
        ('2025-03-14', 'Business Lunch - Nobu',        'Meals',        182.60),
        ('2025-03-16', 'FedEx Shipping',               'Shipping',      29.80),
        ('2025-03-18', 'LinkedIn Premium',             'Subscriptions', 39.99),
        ('2025-03-21', 'Zoom Pro Annual',              'Software',      15.99),
        ('2025-03-23', 'Client Dinner - Spago',        'Meals',        245.30),
        ('2025-03-25', 'Dropbox Business',             'Software',      20.00),
        ('2025-03-28', 'Hertz Car Rental',             'Transport',    175.00),
    ]

    pdf.set_font('Helvetica', '', 9)
    for dt, desc, cat, amt in transactions:
        pdf.cell(30, 7, dt, border=1)
        pdf.cell(80, 7, desc, border=1)
        pdf.cell(25, 7, cat, border=1)
        pdf.cell(30, 7, f'{amt:.2f}', border=1, align='R', ln=True)

    total = sum(t[3] for t in transactions)
    pdf.set_font('Helvetica', 'B', 10)
    pdf.cell(135, 8, 'TOTAL', border=1)
    pdf.cell(30, 8, f'{total:.2f}', border=1, align='R', ln=True)

    out = f'{DESKTOP}/credit_card_statement.pdf'
    pdf.output(out)
    print(f'Created: {out}')


def _create_petty_cash_jpg():
    """Create a petty cash log as a JPG image (handwritten-style table)."""
    try:
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        subprocess.run(['pip3', 'install', 'Pillow'], capture_output=True)
        from PIL import Image, ImageDraw, ImageFont

    W, H = 800, 520
    img = Image.new('RGB', (W, H), color=(255, 253, 245))
    draw = ImageDraw.Draw(img)

    # Title
    draw.rectangle([20, 15, W-20, 55], fill=(200, 220, 255), outline=(50, 80, 160), width=2)
    draw.text((W//2, 35), 'PETTY CASH LOG - March 2025', fill=(20, 40, 120), anchor='mm')

    # Header row
    cols = [20, 120, 360, 480, 620]
    headers = ['Date', 'Description', 'Category', 'Amount', 'Receipt#']
    draw.rectangle([20, 65, W-20, 90], fill=(220, 235, 255), outline=(100, 130, 200), width=1)
    for i, (x, hdr) in enumerate(zip(cols, headers)):
        draw.text((x + 5, 77), hdr, fill=(20, 40, 120), anchor='lm')

    entries = [
        ('03-01', 'Printer paper & pens',       'Office',    42.50,  'PC001'),
        ('03-03', 'Team coffee & snacks',        'Meals',     28.75,  'PC002'),
        ('03-06', 'Taxi to client meeting',      'Transport', 18.00,  'PC003'),
        ('03-10', 'Stamps & mailing supplies',   'Shipping',  12.40,  'PC004'),
        ('03-13', 'Whiteboard markers',          'Office',    15.20,  'PC005'),
        ('03-17', 'Client gift basket',          'Gifts',     65.00,  'PC006'),
        ('03-20', 'Lunch w/ intern candidates',  'Meals',     54.30,  'PC007'),
        ('03-24', 'USB hub & cables',            'Equipment', 34.99,  'PC008'),
        ('03-27', 'Monthly parking pass',        'Transport', 80.00,  'PC009'),
        ('03-30', 'Cleaning supplies office',    'Office',    22.15,  'PC010'),
    ]

    y = 100
    for i, (dt, desc, cat, amt, rcpt) in enumerate(entries):
        bg = (248, 250, 255) if i % 2 == 0 else (255, 255, 255)
        draw.rectangle([20, y, W-20, y+26], fill=bg, outline=(180, 195, 220), width=1)
        draw.text((cols[0]+5, y+13), dt,   fill=(40, 40, 40), anchor='lm')
        draw.text((cols[1]+5, y+13), desc, fill=(40, 40, 40), anchor='lm')
        draw.text((cols[2]+5, y+13), cat,  fill=(40, 40, 40), anchor='lm')
        draw.text((cols[3]+5, y+13), f'${amt:.2f}', fill=(40, 40, 40), anchor='lm')
        draw.text((cols[4]+5, y+13), rcpt, fill=(40, 40, 40), anchor='lm')
        y += 26

    total = sum(e[3] for e in entries)
    draw.rectangle([20, y, W-20, y+28], fill=(200, 220, 255), outline=(50, 80, 160), width=2)
    draw.text((cols[3]+5, y+14), f'TOTAL: ${total:.2f}', fill=(20, 40, 120), anchor='lm')

    out = f'{DESKTOP}/petty_cash_photo.jpg'
    img.save(out, 'JPEG', quality=92)
    print(f'Created: {out}')


def _create_receipt_images():
    """Create 6 realistic receipt images."""
    try:
        from PIL import Image, ImageDraw
    except ImportError:
        from PIL import Image, ImageDraw

    receipts = [
        {
            'file':   'receipt_001.jpg',
            'vendor': 'Starbucks Coffee',
            'date':   '2025-03-03',
            'items':  [('Caffe Latte x2', 9.50), ('Blueberry Muffin', 3.75)],
            'cat':    'Meals',
            'tax_ded': 'Yes',
        },
        {
            'file':   'receipt_002.jpg',
            'vendor': 'Best Buy',
            'date':   '2025-03-08',
            'items':  [('External SSD 1TB', 89.99), ('HDMI Cable', 14.99)],
            'cat':    'Equipment',
            'tax_ded': 'Yes',
        },
        {
            'file':   'receipt_003.jpg',
            'vendor': 'Whole Foods Market',
            'date':   '2025-03-12',
            'items':  [('Client Meeting Catering', 87.45), ('Beverages', 22.10)],
            'cat':    'Meals',
            'tax_ded': 'Yes',
        },
        {
            'file':   'receipt_004.jpg',
            'vendor': 'UPS Store',
            'date':   '2025-03-15',
            'items':  [('Package Shipping Express', 24.60), ('Packing Materials', 8.30)],
            'cat':    'Shipping',
            'tax_ded': 'Yes',
        },
        {
            'file':   'receipt_005.jpg',
            'vendor': 'Walgreens Pharmacy',
            'date':   '2025-03-19',
            'items':  [('First Aid Kit', 18.99), ('Aspirin 200ct', 7.49)],
            'cat':    'Office',
            'tax_ded': 'No',
        },
        {
            'file':   'receipt_006.jpg',
            'vendor': 'Home Depot',
            'date':   '2025-03-26',
            'items':  [('Extension Cord 25ft', 29.97), ('Power Strip 6-outlet', 19.97)],
            'cat':    'Equipment',
            'tax_ded': 'Yes',
        },
    ]

    for r in receipts:
        W, H = 360, 320
        img = Image.new('RGB', (W, H), color=(255, 255, 252))
        draw = ImageDraw.Draw(img)

        # Receipt border
        draw.rectangle([8, 8, W-8, H-8], outline=(160, 160, 160), width=1)

        # Vendor header
        draw.rectangle([8, 8, W-8, 55], fill=(230, 240, 255))
        draw.text((W//2, 30), r['vendor'], fill=(10, 30, 100), anchor='mm')
        draw.text((W//2, 48), f"Date: {r['date']}", fill=(60, 60, 100), anchor='mm')

        # Items
        y = 70
        subtotal = 0.0
        for desc, price in r['items']:
            draw.text((20, y), desc, fill=(40, 40, 40))
            draw.text((W-20, y), f'${price:.2f}', fill=(40, 40, 40), anchor='ra')
            subtotal += price
            y += 26

        draw.line([20, y, W-20, y], fill=(160, 160, 160), width=1)
        y += 8
        tax = round(subtotal * 0.08875, 2)
        total = round(subtotal + tax, 2)
        draw.text((20, y), 'Subtotal:', fill=(40, 40, 40))
        draw.text((W-20, y), f'${subtotal:.2f}', fill=(40, 40, 40), anchor='ra')
        y += 22
        draw.text((20, y), 'Tax (8.875%):', fill=(40, 40, 40))
        draw.text((W-20, y), f'${tax:.2f}', fill=(40, 40, 40), anchor='ra')
        y += 22
        draw.rectangle([16, y-2, W-16, y+22], fill=(220, 235, 220))
        draw.text((20, y+9), 'TOTAL:', fill=(10, 80, 10), anchor='lm')
        draw.text((W-20, y+9), f'${total:.2f}', fill=(10, 80, 10), anchor='rm')
        y += 34

        draw.text((W//2, y+6), f'Category: {r["cat"]}', fill=(80, 80, 120), anchor='mm')
        draw.text((W//2, y+22), f'Tax Deductible: {r["tax_ded"]}', fill=(80, 80, 120), anchor='mm')

        out = f'{DESKTOP}/{r["file"]}'
        img.save(out, 'JPEG', quality=90)
        print(f'Created: {out}')


# ── 3. Open app ────────────────────────────────────────────────────────────────
def open_gui():
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: LibreOffice Calc opened with monthly_report.xlsx (DISPLAY=:0)')


# ── Main ───────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    create_blank_workbook()
    create_desktop_files()
    open_gui()
    print('initial_setup.py complete.')
