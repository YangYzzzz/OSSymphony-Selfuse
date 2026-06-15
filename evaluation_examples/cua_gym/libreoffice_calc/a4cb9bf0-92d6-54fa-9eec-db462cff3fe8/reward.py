"""
Reward Script: Build complete monthly expense report in monthly_report.xlsx
Task ID: osworld_multi_apps_receipt_to_calc_011
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: 'All Transactions' sheet has >= 10 data rows with required columns and a grand total (0.30 pts)
  Component 2: 'By Category' sheet has category breakdown with >= 3 categories and grand total (0.25 pts)
  Component 3: 'Tax Deductible' sheet has filtered business expense rows and a grand total (0.20 pts)
  Component 4: Pie chart exists on 'By Category' sheet (0.25 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_receipt_to_calc_011'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — failure here is a hard gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheet names must exist
    required_sheets = ['All Transactions', 'By Category', 'Tax Deductible']
    for sname in required_sheets:
        if sname not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sname}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # -------------------------------------------------------------------
    # Component 1: 'All Transactions' sheet has >= 10 data rows with
    #   required columns (Date, Description, Category, Amount, Source,
    #   Tax Deductible) and a grand total entry  (0.30 pts)
    # This FAILS on initial (only header row) → PASSES on golden (30 data rows + total)
    # -------------------------------------------------------------------
    try:
        ws_all = wb['All Transactions']

        # Verify the header row has the required columns
        header_row = [ws_all.cell(row=1, column=c).value for c in range(1, 7)]
        required_headers = ['Date', 'Description', 'Category', 'Amount', 'Source', 'Tax Deductible']
        headers_ok = all(h in header_row for h in required_headers)

        if not headers_ok:
            print(f"FAIL: Component 1 — Missing required headers. Found: {header_row}")
        else:
            # Count data rows (rows 2 onwards with non-null Date or Description)
            data_rows = 0
            has_grand_total = False
            for row in ws_all.iter_rows(min_row=2, max_row=ws_all.max_row):
                date_val = row[0].value
                desc_val = row[1].value
                amt_val = row[3].value  # Amount in column D
                if date_val is not None or desc_val is not None:
                    if str(date_val).upper() == 'TOTAL' or str(desc_val).upper().startswith('TOTAL') or str(desc_val).upper().startswith('ALL'):
                        # This looks like a grand total row
                        if amt_val is not None:
                            has_grand_total = True
                    elif amt_val is not None:
                        data_rows += 1

            if data_rows >= 10 and has_grand_total:
                print(f"PASS: Component 1 — 'All Transactions' has {data_rows} data rows with grand total ({0.30} pts)")
                total_score += 0.30
            elif data_rows >= 10:
                print(f"FAIL: Component 1 — 'All Transactions' has {data_rows} data rows but no grand total row found")
            else:
                print(f"FAIL: Component 1 — 'All Transactions' has only {data_rows} data rows (expected >= 10)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: 'By Category' sheet has category breakdown with >= 3
    #   unique categories, each with a total amount, and a grand total (0.25 pts)
    # This FAILS on initial (only header row) → PASSES on golden (10 categories + total)
    # -------------------------------------------------------------------
    try:
        ws_cat = wb['By Category']

        # Count category rows (non-header rows with category name and amount)
        category_rows = 0
        grand_total_found = False
        total_amount = 0.0

        for row in ws_cat.iter_rows(min_row=2, max_row=ws_cat.max_row):
            cat_val = row[0].value
            amt_val = row[1].value
            if cat_val is not None and amt_val is not None:
                if str(cat_val).upper() == 'TOTAL':
                    grand_total_found = True
                    try:
                        total_amount = float(amt_val)
                    except (ValueError, TypeError):
                        pass
                else:
                    category_rows += 1

        if category_rows >= 3 and grand_total_found and total_amount > 0:
            print(f"PASS: Component 2 — 'By Category' has {category_rows} categories, grand total={total_amount} ({0.25} pts)")
            total_score += 0.25
        elif category_rows >= 3:
            print(f"FAIL: Component 2 — 'By Category' has {category_rows} categories but grand total not found or is 0")
        else:
            print(f"FAIL: Component 2 — 'By Category' has only {category_rows} category rows (expected >= 3)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: 'Tax Deductible' sheet has filtered business expense rows
    #   (>= 5 rows) and a grand total row  (0.20 pts)
    # This FAILS on initial (only header row) → PASSES on golden (29 data rows + total)
    # -------------------------------------------------------------------
    try:
        ws_tax = wb['Tax Deductible']

        # Count tax-deductible data rows (not the header, not the total row)
        tax_rows = 0
        tax_grand_total = False

        for row in ws_tax.iter_rows(min_row=2, max_row=ws_tax.max_row):
            date_val = row[0].value
            desc_val = row[1].value
            amt_val = row[3].value  # Amount in column D
            if date_val is not None or desc_val is not None:
                if str(date_val).upper() == 'TOTAL' or (desc_val and str(desc_val).upper().startswith('TAX')):
                    if amt_val is not None:
                        tax_grand_total = True
                elif amt_val is not None:
                    tax_rows += 1

        if tax_rows >= 5 and tax_grand_total:
            print(f"PASS: Component 3 — 'Tax Deductible' has {tax_rows} rows with grand total ({0.20} pts)")
            total_score += 0.20
        elif tax_rows >= 5:
            print(f"FAIL: Component 3 — 'Tax Deductible' has {tax_rows} rows but no grand total row found")
        else:
            print(f"FAIL: Component 3 — 'Tax Deductible' has only {tax_rows} rows (expected >= 5)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: A pie chart exists on the 'By Category' sheet
    #   visualizing the category breakdown  (0.25 pts)
    # This FAILS on initial (no charts) → PASSES on golden (1 pie chart)
    # -------------------------------------------------------------------
    try:
        ws_cat_chart = wb['By Category']
        charts = ws_cat_chart._charts

        if len(charts) >= 1:
            # Verify at least one of them is a pie chart
            from openpyxl.chart import PieChart
            pie_charts = [c for c in charts if isinstance(c, PieChart)]
            if pie_charts:
                print(f"PASS: Component 4 — Pie chart found on 'By Category' sheet ({0.25} pts)")
                total_score += 0.25
            else:
                # Accept any chart type as partial credit is not possible here,
                # but still award points if a chart exists (category breakdown chart)
                chart_types = [type(c).__name__ for c in charts]
                print(f"FAIL: Component 4 — Charts found on 'By Category' but none are PieChart: {chart_types}")
        else:
            print(f"FAIL: Component 4 — No charts found on 'By Category' sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/monthly_report.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
