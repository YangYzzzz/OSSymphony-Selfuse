"""
Initial Setup: Address Book and Mailing List Manager
Task ID: calc_wf_095
Domain: libreoffice_calc

Creates a workbook with 45 contacts, data validation for states and zip codes,
conditional formatting for inactive contacts. Labels and Summary sheets exist
but are empty (task requires the agent to populate them).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

US_STATES = [
    'AL','AK','AZ','AR','CA','CO','CT','DE','FL','GA',
    'HI','ID','IL','IN','IA','KS','KY','LA','ME','MD',
    'MA','MI','MN','MS','MO','MT','NE','NV','NH','NJ',
    'NM','NY','NC','ND','OH','OK','OR','PA','RI','SC',
    'SD','TN','TX','UT','VT','VA','WA','WV','WI','WY'
]

CONTACTS = [
    ('Sarah', 'Chen', '1425 Maple Ave', 'San Francisco', 'CA', '94117', 'Business', 'Active'),
    ('Marcus', 'Johnson', '782 Oak Blvd', 'Austin', 'TX', '78701', 'Friends', 'Active'),
    ('Emily', 'Rodriguez', '330 Pine St', 'Denver', 'CO', '80202', 'Family', 'Active'),
    ('James', 'Williams', '9501 Elm Dr', 'Seattle', 'WA', '98101', 'Business', 'Active'),
    ('Olivia', 'Martinez', '217 Cedar Ln', 'Portland', 'OR', '97201', 'Holiday', 'Active'),
    ('Daniel', 'Brown', '4488 Birch Rd', 'Chicago', 'IL', '60601', 'Friends', 'Inactive'),
    ('Sophia', 'Lee', '5612 Spruce Way', 'Boston', 'MA', '02108', 'Family', 'Active'),
    ('William', 'Davis', '893 Walnut Ct', 'Phoenix', 'AZ', '85001', 'Business', 'Active'),
    ('Isabella', 'Garcia', '1076 Ash St', 'Miami', 'FL', '33101', 'Holiday', 'Active'),
    ('Benjamin', 'Wilson', '2345 Cherry Dr', 'Nashville', 'TN', '37201', 'Friends', 'Active'),
    ('Mia', 'Anderson', '678 Willow Ln', 'Atlanta', 'GA', '30301', 'Family', 'Inactive'),
    ('Ethan', 'Thomas', '4321 Hickory Ave', 'Dallas', 'TX', '75201', 'Business', 'Active'),
    ('Charlotte', 'Taylor', '159 Magnolia Blvd', 'Minneapolis', 'MN', '55401', 'Holiday', 'Active'),
    ('Alexander', 'Moore', '8765 Poplar St', 'Las Vegas', 'NV', '89101', 'Friends', 'Active'),
    ('Amelia', 'Jackson', '2109 Sycamore Rd', 'Charlotte', 'NC', '28201', 'Family', 'Active'),
    ('Henry', 'White', '3456 Redwood Ct', 'San Diego', 'CA', '92101', 'Business', 'Inactive'),
    ('Harper', 'Harris', '7890 Juniper Way', 'Columbus', 'OH', '43201', 'Holiday', 'Active'),
    ('Sebastian', 'Clark', '1234 Dogwood Dr', 'Indianapolis', 'IN', '46201', 'Friends', 'Active'),
    ('Evelyn', 'Lewis', '5678 Hawthorn Ln', 'Jacksonville', 'FL', '32099', 'Family', 'Active'),
    ('Jack', 'Robinson', '9012 Chestnut Ave', 'San Antonio', 'TX', '78201', 'Business', 'Active'),
    ('Abigail', 'Walker', '3456 Cottonwood Blvd', 'Philadelphia', 'PA', '19101', 'Holiday', 'Inactive'),
    ('Owen', 'Hall', '7890 Cypress St', 'Detroit', 'MI', '48201', 'Friends', 'Active'),
    ('Ella', 'Allen', '2345 Mulberry Rd', 'Memphis', 'TN', '38101', 'Family', 'Active'),
    ('Luke', 'Young', '6789 Beech Ct', 'Baltimore', 'MD', '21201', 'Business', 'Active'),
    ('Scarlett', 'King', '1234 Laurel Way', 'Milwaukee', 'WI', '53201', 'Holiday', 'Active'),
    ('Gabriel', 'Wright', '5678 Aspen Dr', 'Albuquerque', 'NM', '87101', 'Friends', 'Inactive'),
    ('Chloe', 'Lopez', '9012 Basswood Ln', 'Tucson', 'AZ', '85701', 'Family', 'Active'),
    ('Carter', 'Hill', '3456 Catalpa Ave', 'Fresno', 'CA', '93701', 'Business', 'Active'),
    ('Aria', 'Scott', '7890 Alder Blvd', 'Sacramento', 'CA', '95801', 'Holiday', 'Active'),
    ('Jayden', 'Green', '2345 Linden St', 'Kansas City', 'MO', '64101', 'Friends', 'Active'),
    ('Grace', 'Adams', '6789 Hemlock Rd', 'Mesa', 'AZ', '85201', 'Family', 'Active'),
    ('Lincoln', 'Baker', '1234 Ironwood Ct', 'Omaha', 'NE', '68101', 'Business', 'Inactive'),
    ('Lily', 'Nelson', '5678 Tamarack Way', 'Raleigh', 'NC', '27601', 'Holiday', 'Active'),
    ('Mateo', 'Carter', '9012 Boxelder Dr', 'Virginia Beach', 'VA', '23450', 'Friends', 'Active'),
    ('Zoey', 'Mitchell', '3456 Butternut Ln', 'Colorado Springs', 'CO', '80901', 'Family', 'Active'),
    ('Levi', 'Perez', '7890 Hackberry Ave', 'Tampa', 'FL', '33601', 'Business', 'Active'),
    ('Hannah', 'Roberts', '2345 Sweetgum Blvd', 'St. Louis', 'MO', '63101', 'Holiday', 'Inactive'),
    ('Julian', 'Turner', '6789 Buckeye St', 'Pittsburgh', 'PA', '15201', 'Friends', 'Active'),
    ('Nora', 'Phillips', '1234 Pawpaw Rd', 'Cincinnati', 'OH', '45201', 'Family', 'Active'),
    ('Grayson', 'Campbell', '5678 Sassafras Ct', 'Orlando', 'FL', '32801', 'Business', 'Active'),
    ('Riley', 'Parker', '9012 Tupelo Way', 'New Orleans', 'LA', '70112', 'Holiday', 'Active'),
    ('Elijah', 'Evans', '3456 Sourwood Dr', 'Cleveland', 'OH', '44101', 'Friends', 'Active'),
    ('Layla', 'Edwards', '7890 Serviceberry Ln', 'Honolulu', 'HI', '96801', 'Family', 'Inactive'),
    ('Nathan', 'Collins', '2345 Redbud Ave', 'Salt Lake City', 'UT', '84101', 'Business', 'Active'),
    ('Penelope', 'Stewart', '6789 Sumac Blvd', 'Richmond', 'VA', '23219', 'Holiday', 'Active'),
]

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Contacts ---
    ws = wb.active
    ws.title = 'Contacts'

    headers = ['First Name', 'Last Name', 'Address', 'City', 'State', 'Zip', 'Category', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Write contact data
    for r, contact in enumerate(CONTACTS, 2):
        for c, val in enumerate(contact, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 26
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Data validation: State abbreviations (column E)
    state_list = ','.join(US_STATES)
    dv_state = DataValidation(
        type='list',
        formula1=f'"{state_list}"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_state.error = 'Please enter a valid US state abbreviation'
    dv_state.errorTitle = 'Invalid State'
    dv_state.prompt = 'Select a US state abbreviation'
    dv_state.promptTitle = 'State'
    dv_state.add('E2:E100')
    ws.add_data_validation(dv_state)

    # Data validation: Zip code (5-digit format, column F)
    dv_zip = DataValidation(
        type='textLength',
        operator='equal',
        formula1='5',
        allow_blank=False,
        showDropDown=False,
    )
    dv_zip.error = 'Zip code must be exactly 5 digits'
    dv_zip.errorTitle = 'Invalid Zip Code'
    dv_zip.prompt = 'Enter a 5-digit zip code'
    dv_zip.promptTitle = 'Zip Code'
    dv_zip.add('F2:F100')
    ws.add_data_validation(dv_zip)

    # Data validation: Category (column G)
    dv_cat = DataValidation(
        type='list',
        formula1='"Family,Friends,Business,Holiday"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_cat.error = 'Please select a valid category'
    dv_cat.errorTitle = 'Invalid Category'
    dv_cat.add('G2:G100')
    ws.add_data_validation(dv_cat)

    # Data validation: Status (column H)
    dv_status = DataValidation(
        type='list',
        formula1='"Active,Inactive"',
        allow_blank=False,
        showDropDown=False,
    )
    dv_status.error = 'Please select Active or Inactive'
    dv_status.errorTitle = 'Invalid Status'
    dv_status.add('H2:H100')
    ws.add_data_validation(dv_status)

    # Conditional formatting: gray fill for inactive contacts
    gray_fill = PatternFill(start_color='FFD9D9D9', end_color='FFD9D9D9', fill_type='solid')
    gray_font = Font(color='808080')
    ws.conditional_formatting.add(
        'A2:H46',
        FormulaRule(
            formula=['$H2="Inactive"'],
            fill=gray_fill,
            font=gray_font,
        )
    )

    # Auto-filter on Contacts
    ws.auto_filter.ref = 'A1:H46'

    # --- Sheet 2: Labels (empty, to be populated by agent) ---
    ws_labels = wb.create_sheet('Labels')
    ws_labels['A1'] = 'Avery 5160 Label Layout'
    ws_labels['A1'].font = Font(size=14, bold=True, color='333333')
    ws_labels['A2'] = '(3 columns x 10 rows per page - 30 labels per page)'
    ws_labels['A2'].font = Font(size=10, italic=True, color='666666')

    # --- Sheet 3: Summary (empty, to be populated by agent) ---
    ws_summary = wb.create_sheet('Summary')
    ws_summary['A1'] = 'Mailing List Summary'
    ws_summary['A1'].font = Font(size=14, bold=True, color='333333')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
