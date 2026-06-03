"""
Reward Script: Address Book and Mailing List Manager
Task ID: calc_wf_095
Domain: libreoffice_calc
Scoring:
  Component 1: Summary sheet has COUNTIFS formulas for category x status (0.25)
  Component 2: Summary sheet has mailing statistics (active count, stamp rate, postage) (0.20)
  Component 3: Summary sheet has a pie chart (0.15)
  Component 4: Labels sheet has 3-column layout with INDEX formulas for active contacts (0.30)
  Component 5: Labels sheet has content spanning multiple label rows (at least 10 rows of labels) (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_095'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Required sheets must exist
    required_sheets = ['Contacts', 'Labels', 'Summary']
    for sname in required_sheets:
        if sname not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sname}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    ws_contacts = wb['Contacts']
    ws_labels = wb['Labels']
    ws_summary = wb['Summary']

    # =========================================================================
    # Component 1: Summary sheet has COUNTIFS formulas for category x status
    # (0.25 points)
    # The golden has COUNTIFS in B4:C7 for Family/Friends/Business/Holiday
    # and SUM totals in B8:D8. Initial Summary has only a header in A1.
    # =========================================================================
    try:
        countifs_found = 0
        categories_found = []
        expected_categories = {'Family', 'Friends', 'Business', 'Holiday'}

        for row_num in range(2, ws_summary.max_row + 1):
            cat_val = ws_summary.cell(row=row_num, column=1).value
            b_val = ws_summary.cell(row=row_num, column=2).value
            c_val = ws_summary.cell(row=row_num, column=3).value

            if cat_val and str(cat_val).strip() in expected_categories:
                categories_found.append(str(cat_val).strip())
                # Check if B and C columns have COUNTIFS formulas
                if b_val and isinstance(b_val, str) and 'COUNTIFS' in b_val.upper():
                    countifs_found += 1
                if c_val and isinstance(c_val, str) and 'COUNTIFS' in c_val.upper():
                    countifs_found += 1

        # Need at least 4 categories with COUNTIFS (at least in Active column)
        cats_with_formulas = len(set(categories_found))
        if cats_with_formulas >= 4 and countifs_found >= 4:
            print(f"PASS: Component 1 -- Summary has COUNTIFS for {cats_with_formulas} categories, {countifs_found} formulas found (0.25 pts)")
            total_score += 0.25
        elif cats_with_formulas >= 2 and countifs_found >= 2:
            print(f"PARTIAL: Component 1 -- Summary has COUNTIFS for {cats_with_formulas} categories, {countifs_found} formulas (0.125 pts)")
            total_score += 0.125
        else:
            print(f"FAIL: Component 1 -- Expected COUNTIFS for 4 categories, found {cats_with_formulas} categories with {countifs_found} formulas")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Component 2: Summary sheet has mailing statistics
    # (0.20 points)
    # Active mailing count (COUNTIF), stamp rate ($0.63), postage estimate formula
    # Initial Summary has none of these.
    # =========================================================================
    try:
        active_count_found = 0
        stamp_rate_found = 0
        postage_formula_found = 0

        for row_num in range(2, ws_summary.max_row + 1):
            label = ws_summary.cell(row=row_num, column=1).value
            b_val = ws_summary.cell(row=row_num, column=2).value

            if label and 'active' in str(label).lower() and 'count' in str(label).lower():
                # Check for COUNTIF formula
                if b_val and isinstance(b_val, str) and 'COUNTIF' in b_val.upper():
                    active_count_found = 1

            if label and 'stamp' in str(label).lower() and 'rate' in str(label).lower():
                # Check for 0.63 value
                if b_val is not None:
                    try:
                        if abs(float(b_val) - 0.63) < 0.01:
                            stamp_rate_found = 1
                    except (ValueError, TypeError):
                        pass

            if label and 'postage' in str(label).lower():
                # Check for formula referencing active count and rate
                if b_val and isinstance(b_val, str) and '=' in str(b_val):
                    postage_formula_found = 1

        sub_score = active_count_found * 0.08 + stamp_rate_found * 0.06 + postage_formula_found * 0.06

        if sub_score > 0:
            print(f"PASS: Component 2 -- Mailing stats: active_count={active_count_found}, stamp_rate={stamp_rate_found}, postage={postage_formula_found} ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 2 -- No mailing statistics found in Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Summary sheet has a pie chart
    # (0.15 points)
    # Initial Summary has no charts. Golden has 1 PieChart.
    # =========================================================================
    try:
        charts = ws_summary._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_type = type(chart).__name__
            if 'Pie' in chart_type:
                print(f"PASS: Component 3 -- Summary has PieChart ({chart_type}) (0.15 pts)")
                total_score += 0.15
            elif len(charts) >= 1:
                # Any chart is partial credit
                print(f"PARTIAL: Component 3 -- Summary has chart but type is {chart_type}, expected PieChart (0.08 pts)")
                total_score += 0.08
        else:
            print(f"FAIL: Component 3 -- No charts found in Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================================
    # Component 4: Labels sheet has 3-column layout with INDEX formulas
    # (0.30 points)
    # Golden has formulas in columns A, D, G using INDEX/SMALL/IF to pull
    # active contacts. Initial Labels has only title/subtitle.
    # =========================================================================
    try:
        # Check for formulas in the 3-column layout (cols A, D, G)
        formula_cols = {1: 0, 4: 0, 7: 0}  # col_num -> count of formula cells
        index_formula_count = 0

        for row_num in range(3, ws_labels.max_row + 1):
            for col_num in [1, 4, 7]:
                val = ws_labels.cell(row=row_num, column=col_num).value
                if val and isinstance(val, str) and '=' in val:
                    formula_cols[col_num] += 1
                    if 'INDEX' in val.upper():
                        index_formula_count += 1

        total_formulas = sum(formula_cols.values())
        cols_with_formulas = sum(1 for c in formula_cols.values() if c > 0)

        if cols_with_formulas >= 3 and index_formula_count >= 6:
            print(f"PASS: Component 4 -- Labels has 3-column layout with INDEX formulas: col_A={formula_cols[1]}, col_D={formula_cols[4]}, col_G={formula_cols[7]}, INDEX count={index_formula_count} (0.30 pts)")
            total_score += 0.30
        elif cols_with_formulas >= 2 and index_formula_count >= 3:
            print(f"PARTIAL: Component 4 -- Labels has partial layout: {cols_with_formulas} columns with formulas, {index_formula_count} INDEX formulas (0.15 pts)")
            total_score += 0.15
        elif total_formulas > 0:
            print(f"PARTIAL: Component 4 -- Labels has some formulas but incomplete layout: {total_formulas} total formulas (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 -- No formulas found in Labels sheet layout columns")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =========================================================================
    # Component 5: Labels sheet has content spanning multiple label rows
    # (0.10 points)
    # Golden has labels from row 4 to row 42 (at least 10 rows of label blocks).
    # Initial has only rows 1-2 with title/subtitle.
    # =========================================================================
    try:
        # Count rows with actual content (formulas or values) beyond row 2
        content_rows = 0
        for row_num in range(3, ws_labels.max_row + 1):
            row_values = [ws_labels.cell(row=row_num, column=col_num).value
                          for col_num in range(1, ws_labels.max_column + 1)]
            if any(v is not None for v in row_values):
                content_rows += 1

        if content_rows >= 10:
            print(f"PASS: Component 5 -- Labels has {content_rows} content rows beyond header (0.10 pts)")
            total_score += 0.10
        elif content_rows >= 3:
            print(f"PARTIAL: Component 5 -- Labels has {content_rows} content rows (expected >= 10) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 -- Labels has only {content_rows} content rows beyond header")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
