"""
Initial Setup: Statistical outlier spreadsheet with measurement data
Task ID: calc_gcv_043
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Statistical_Outliers"

    # Headers
    headers = ["Sample ID", "Batch", "Lab", "Technician", "Method", "Measurement"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data
    batches = ["B-2025-01", "B-2025-02", "B-2025-03", "B-2025-04", "B-2025-05"]
    labs = ["Central Lab", "West Wing Lab", "East Annex", "Satellite Lab A", "Satellite Lab B"]
    technicians = [
        "Sarah Chen", "Marcus Johnson", "Elena Rodriguez", "David Kim",
        "Priya Patel", "James O'Brien", "Aisha Mohammed", "Lucas Weber",
        "Mei-Ling Wu", "Carlos Gutierrez"
    ]
    methods = ["HPLC", "GC-MS", "UV-Vis", "ICP-OES", "AAS"]

    random.seed(42)

    # Generate 54 rows of data with mostly normal distribution (mean=50, SD=10)
    # but inject some extreme outliers
    measurements = []
    for i in range(54):
        if i in [5, 18, 31, 47]:  # outlier positions (beyond 2 SD)
            # Extreme values: > 70 or < 30 (more than 2 SD from mean 50)
            outlier_vals = [15.3, 82.7, 9.1, 88.4]
            measurements.append(outlier_vals[len([x for x in [5, 18, 31, 47] if x <= i]) - 1])
        elif i in [10, 25, 40]:  # borderline outliers (close to 2 SD)
            borderline_vals = [29.5, 71.2, 28.8]
            measurements.append(borderline_vals[len([x for x in [10, 25, 40] if x <= i]) - 1])
        else:
            # Normal distribution around mean=50, SD=10, clamped to avoid accidental outliers
            val = random.gauss(50, 8)
            val = max(32, min(68, val))  # keep within ~2 SD
            measurements.append(round(val, 1))

    for i in range(54):
        row = i + 2
        sample_id = f"SMP-{2025000 + i + 1:07d}"
        batch = batches[i % len(batches)]
        lab = labs[i % len(labs)]
        tech = technicians[i % len(technicians)]
        method = methods[i % len(methods)]
        measurement = measurements[i]

        ws.cell(row=row, column=1, value=sample_id)
        ws.cell(row=row, column=2, value=batch)
        ws.cell(row=row, column=3, value=lab)
        ws.cell(row=row, column=4, value=tech)
        ws.cell(row=row, column=5, value=method)
        ws.cell(row=row, column=6, value=measurement)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
