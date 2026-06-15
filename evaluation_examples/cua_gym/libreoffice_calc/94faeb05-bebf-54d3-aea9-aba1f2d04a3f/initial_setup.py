"""
Initial Setup: SUMIF transaction data spreadsheet
Task ID: calc_fmb_sumif_numeric_criteria_049
Domain: libreoffice_calc
"""

import openpyxl
import random
import datetime
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sumif_numeric_criteria_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # --- Headers ---
    headers = ['Trans ID', 'Date', 'Amount', 'Type', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Style headers bold
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # --- Generate amounts ---
    # Need exactly 87 records > 1000 summing to 318450
    # And 213 records <= 1000 (ranging 12.50 to 999.99)

    random.seed(42)

    # Build 87 high-value amounts (> 1000) summing to exactly 318450
    high_value_amounts = [
        9800.00, 9250.50, 8900.00, 8750.25, 8500.00,
        8200.75, 7980.00, 7650.50, 7400.00, 7100.25,
        6850.00, 6600.75, 6350.00, 6100.50, 5900.00,
        5650.75, 5400.00, 5200.25, 4950.00, 4700.50,
        4500.00, 4300.25, 4100.00, 3950.75, 3800.00,
        3650.50, 3500.00, 3350.25, 3200.00, 3050.75,
        2900.00, 2780.50, 2650.00, 2520.25, 2400.00,
        2290.75, 2180.00, 2070.50, 1960.00, 1860.25,
        1760.00, 1670.75, 1580.00, 1500.50, 1425.00,
        1350.25, 1290.00, 1230.75, 1180.00, 1140.50,
        1100.00, 1070.25, 1050.00, 1030.75, 1020.00,
        1015.50, 1012.00, 1010.25, 1008.00, 1006.75,
        1005.00, 1004.50, 1003.00, 1002.75, 1002.00,
        1001.75, 1001.50, 1001.25, 1001.00, 1000.75,
    ]

    # Filter to only > 1000
    high_value_amounts = [a for a in high_value_amounts if a > 1000.0]

    # Fill up to 86 entries (leaving slot for the balancing value)
    idx = 0
    fill_values = [1001.00 + idx * 50 for idx in range(100)]
    while len(high_value_amounts) < 86:
        high_value_amounts.append(fill_values[idx])
        idx += 1

    # Add the balancing 87th value
    last_val = round(318450.00 - sum(high_value_amounts), 2)
    assert last_val > 1000.0, f"Balancing value {last_val} not > 1000"
    high_value_amounts.append(last_val)

    assert len(high_value_amounts) == 87
    assert all(a > 1000.0 for a in high_value_amounts)
    assert round(sum(high_value_amounts), 2) == 318450.00

    # Build 213 low-value amounts (<= 1000, ranging 12.50 to 999.99)
    # Generate varied realistic transaction amounts
    low_value_amounts = []

    # Large low-value transactions (500-999.99): 60 records
    for i in range(60):
        val = round(500.00 + (i * 8.33), 2)
        if val > 999.99:
            val = 999.99 - i * 0.01
        low_value_amounts.append(val)

    # Medium low-value transactions (100-499.99): 80 records
    for i in range(80):
        val = round(100.00 + (i * 5.00), 2)
        if val > 499.99:
            val = 499.99 - i * 0.01
        low_value_amounts.append(val)

    # Small transactions (12.50-99.99): 73 records
    for i in range(73):
        val = round(12.50 + (i * 1.20), 2)
        if val > 99.99:
            val = 12.50 + (i % 7) * 12.50
        low_value_amounts.append(val)

    assert len(low_value_amounts) == 213, f"Got {len(low_value_amounts)} low-value amounts"
    assert all(0 < a <= 1000.0 for a in low_value_amounts), \
        [a for a in low_value_amounts if a > 1000.0 or a <= 0]

    # Combine and shuffle
    all_amounts = high_value_amounts + low_value_amounts
    random.shuffle(all_amounts)
    assert len(all_amounts) == 300

    # Transaction types and statuses
    types = ['Purchase', 'Refund', 'Transfer', 'Payment', 'Withdrawal', 'Deposit']
    statuses = ['Completed', 'Pending', 'Failed', 'Approved', 'Processing']
    trans_id_prefixes = ['TXN', 'ORD', 'PMT', 'REF', 'TRF']

    # Dates: 2024-01-02 onwards
    start_date = datetime.date(2024, 1, 2)
    dates = [(start_date + datetime.timedelta(days=i)).strftime('%Y-%m-%d')
             for i in range(300)]

    # Write 300 rows (rows 2-301)
    for i, amount in enumerate(all_amounts):
        row = i + 2
        trans_id = f'{trans_id_prefixes[i % len(trans_id_prefixes)]}-{2024000 + i + 1:07d}'
        date_val = dates[i]
        t_type = types[i % len(types)]
        # Row 2: E2 = 'High Value Total' (as per context)
        status = 'High Value Total' if row == 2 else statuses[i % len(statuses)]

        ws.cell(row=row, column=1, value=trans_id)
        ws.cell(row=row, column=2, value=date_val)
        ws.cell(row=row, column=3, value=amount)
        ws.cell(row=row, column=4, value=t_type)
        ws.cell(row=row, column=5, value=status)
        # Column F: empty for all rows (F2 is the target)

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Transactions | 1 header + 300 data rows (rows 2-301)')
    print(f'High-value records (>$1,000): 87 | Sum: {sum(high_value_amounts):.2f}')
    print(f'Low-value records (<=1000): 213')
    print(f'E2 = "High Value Total", F2 = empty (target for SUMIF formula)')


create_initial()
