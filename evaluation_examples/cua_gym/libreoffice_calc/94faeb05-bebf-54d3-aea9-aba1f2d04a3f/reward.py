"""
Reward Script: Sum high-value transactions using SUMIF in cell F2
Task ID: calc_fmb_sumif_numeric_criteria_049
Domain: libreoffice_calc

Scoring Rubric:
  Precondition gate: 'Transactions' sheet exists with headers and data intact (no points, early exit if violated)
  Component 1: F2 contains a SUMIF formula (0.5 pts)
  Component 2: SUMIF formula in F2 references C2:C301 with ">1000" criteria exactly (0.5 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_sumif_numeric_criteria_049'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Transactions' sheet must exist with correct structure
    # This is a gate (no points) — if violated, task is fundamentally broken
    if 'Transactions' not in wb.sheetnames:
        print("FAIL GATE: 'Transactions' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Transactions']

    # Precondition gate: Verify headers and data range are intact
    expected_headers = ['Trans ID', 'Date', 'Amount', 'Type', 'Status']
    actual_headers = [ws.cell(row=1, column=col).value for col in range(1, 6)]
    if actual_headers != expected_headers:
        print(f"FAIL GATE: Headers modified. Expected {expected_headers}, found {actual_headers}")
        print("REWARD: 0.0")
        return 0.0

    if ws.max_row < 301:
        print(f"FAIL GATE: Data rows missing. Expected >= 301, found {ws.max_row}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: F2 contains a SUMIF formula (0.5 points)
    # In the initial file, F2 is None/empty. The task requires placing a SUMIF formula in F2.
    try:
        f2_value = ws['F2'].value
        if f2_value is not None and isinstance(f2_value, str) and f2_value.upper().startswith('=SUMIF'):
            print(f"PASS: Component 1 — F2 contains a SUMIF formula: {repr(f2_value)} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — F2 should contain a SUMIF formula, found: {repr(f2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: SUMIF formula in F2 references C2:C301 with criteria ">1000" (0.5 points)
    # The formula must specifically target the correct data range and threshold.
    try:
        f2_value = ws['F2'].value
        if f2_value is not None and isinstance(f2_value, str) and f2_value.upper().startswith('=SUMIF'):
            formula_normalized = f2_value.upper().replace(' ', '')
            # Check that the range C2:C301 is referenced
            has_correct_range = 'C2:C301' in formula_normalized
            # Check that the criteria ">1000" is present (handle various quote styles)
            has_correct_criteria = ('">1000"' in f2_value) or ("'>1000'" in f2_value)
            if has_correct_range and has_correct_criteria:
                print(f"PASS: Component 2 — SUMIF references C2:C301 with '>1000' criteria (0.5 pts)")
                total_score += 0.5
            elif not has_correct_range:
                print(f"FAIL: Component 2 — Expected range 'C2:C301' in SUMIF formula, found: {repr(f2_value)}")
            else:
                print(f"FAIL: Component 2 — Expected criteria '>1000' (with quotes) in SUMIF formula, found: {repr(f2_value)}")
        else:
            print(f"FAIL: Component 2 — F2 is empty or not a SUMIF formula: {repr(f2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
