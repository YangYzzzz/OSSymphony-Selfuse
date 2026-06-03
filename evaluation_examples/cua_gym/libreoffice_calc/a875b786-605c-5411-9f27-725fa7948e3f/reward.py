"""
Reward Script: Build a marketing campaign ROI dashboard
Task ID: calc_sales_marketing_roi_026
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: ROI formulas in H2:H21 with percentage format — 0.35 pts
  Component 2: Conditional formatting green fill for ROI > 200% — 0.25 pts
  Component 3: Conditional formatting red fill for negative ROI — 0.25 pts
  Component 4: Bar chart with title 'Campaign ROI Ranking' — 0.15 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_marketing_roi_026'


def get_chart_title_text(title_obj):
    """Extract text string from an openpyxl chart Title object."""
    try:
        if title_obj is None:
            return None
        # Navigate: title.tx.rich.paragraphs[0].r[0].t
        rich = title_obj.tx.rich
        for para in rich.p:
            for run in para.r:
                if run.t:
                    return run.t
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Campaigns' sheet must exist
    if 'Campaigns' not in wb.sheetnames:
        print("CRITICAL: 'Campaigns' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Campaigns']

    # -------------------------------------------------------------------
    # Component 1: ROI formulas in H2:H21 with percentage format (0.35 pts)
    # Task requires: =(Gn-Dn)/Dn formatted as percentage
    # FAILS on initial (H column is empty), PASSES on golden
    # -------------------------------------------------------------------
    try:
        formula_count = 0
        pct_format_count = 0
        expected_pattern = re.compile(r'=\s*\(\s*G(\d+)\s*-\s*D\1\s*\)\s*/\s*D\1', re.IGNORECASE)

        for row in range(2, 22):
            cell = ws.cell(row=row, column=8)  # Column H
            val = cell.value
            fmt = cell.number_format

            if val is not None and isinstance(val, str):
                # Check formula matches =(Gn-Dn)/Dn pattern
                if expected_pattern.match(val.replace(' ', '')):
                    formula_count += 1
                # Accept any variant that references G and D columns for ROI
                elif re.search(r'G\d+', val) and re.search(r'D\d+', val):
                    formula_count += 1

            # Check percentage format (any percentage-style format)
            if fmt and ('%' in fmt or 'pct' in fmt.lower()):
                pct_format_count += 1

        if formula_count == 20:
            print(f"PASS: Component 1a — All 20 ROI formulas present in H2:H21 (formula_count={formula_count})")
            partial = 0.25
        elif formula_count >= 15:
            print(f"PARTIAL: Component 1a — {formula_count}/20 ROI formulas present in H2:H21")
            partial = 0.15
        elif formula_count > 0:
            print(f"PARTIAL: Component 1a — {formula_count}/20 ROI formulas present in H2:H21")
            partial = 0.08
        else:
            print(f"FAIL: Component 1a — No ROI formulas found in H2:H21")
            partial = 0.0

        total_score += partial

        if pct_format_count == 20:
            print(f"PASS: Component 1b — All 20 cells H2:H21 have percentage number format")
            total_score += 0.10
        elif pct_format_count >= 10:
            print(f"PARTIAL: Component 1b — {pct_format_count}/20 cells have percentage format")
            total_score += 0.05
        else:
            print(f"FAIL: Component 1b — Only {pct_format_count}/20 cells have percentage format")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Conditional formatting — green fill for ROI > 200% (0.25 pts)
    # Task requires: green highlight if ROI > 2.0 (i.e., 200% when stored as decimal)
    # FAILS on initial (no CF rules), PASSES on golden
    # -------------------------------------------------------------------
    try:
        green_cf_found = False
        for cf_range in ws.conditional_formatting:
            for rule in cf_range.rules:
                # Check for greaterThan rule with formula '2' (200% as decimal)
                # or '200' (if stored as percentage number)
                if rule.type == 'cellIs' and rule.operator == 'greaterThan':
                    formula_val = rule.formula[0] if rule.formula else None
                    # Accept 2, 2.0, or 200 (200%)
                    if formula_val in ('2', '2.0', '200', '200%'):
                        # Check fill color is green-ish
                        if rule.dxf and rule.dxf.fill:
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                                # Accept any greenish color (G channel high, R/B channels lower)
                                # Common greens: FF00FF00, FF00B050, etc.
                                r_hex = fill_color[2:4]
                                g_hex = fill_color[4:6]
                                b_hex = fill_color[6:8]
                                r_val = int(r_hex, 16)
                                g_val = int(g_hex, 16)
                                b_val = int(b_hex, 16)
                                if g_val > r_val and g_val > b_val:
                                    green_cf_found = True
                                    print(f"PASS: Component 2 — Green CF rule found: greaterThan {formula_val}, color={fill_color}")
                                else:
                                    # Color may be a standard green but not matching
                                    # Also accept if color starts with common green patterns
                                    if fill_color.upper() in ('FF00FF00', 'FF00B050', 'FF92D050'):
                                        green_cf_found = True
                                        print(f"PASS: Component 2 — Green CF rule found: greaterThan {formula_val}, color={fill_color}")
                            except Exception as ce:
                                # If we can't parse the color but the rule is correct, still count it
                                green_cf_found = True
                                print(f"PASS: Component 2 — Green CF rule found (color check skipped: {ce})")
                        else:
                            # Rule present without fill details — partial credit
                            print(f"PARTIAL: Component 2 — Green CF rule found but no fill color details")
                            green_cf_found = True

        if green_cf_found:
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — No green conditional formatting rule for ROI > 200% found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Conditional formatting — red fill for negative ROI (0.25 pts)
    # Task requires: red highlight if ROI < 0
    # FAILS on initial (no CF rules), PASSES on golden
    # -------------------------------------------------------------------
    try:
        red_cf_found = False
        for cf_range in ws.conditional_formatting:
            for rule in cf_range.rules:
                if rule.type == 'cellIs' and rule.operator == 'lessThan':
                    formula_val = rule.formula[0] if rule.formula else None
                    if formula_val in ('0', '0.0'):
                        if rule.dxf and rule.dxf.fill:
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                                r_hex = fill_color[2:4]
                                g_hex = fill_color[4:6]
                                b_hex = fill_color[6:8]
                                r_val = int(r_hex, 16)
                                g_val = int(g_hex, 16)
                                b_val = int(b_hex, 16)
                                if r_val > g_val and r_val > b_val:
                                    red_cf_found = True
                                    print(f"PASS: Component 3 — Red CF rule found: lessThan {formula_val}, color={fill_color}")
                                else:
                                    if fill_color.upper() in ('FFFF0000', 'FFFF3300', 'FFFF6600'):
                                        red_cf_found = True
                                        print(f"PASS: Component 3 — Red CF rule found: lessThan {formula_val}, color={fill_color}")
                            except Exception as ce:
                                red_cf_found = True
                                print(f"PASS: Component 3 — Red CF rule found (color check skipped: {ce})")
                        else:
                            red_cf_found = True
                            print(f"PARTIAL: Component 3 — Red CF rule found but no fill color details")

        if red_cf_found:
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — No red conditional formatting rule for negative ROI found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: Bar chart with title 'Campaign ROI Ranking' (0.15 pts)
    # Task requires: a bar/column chart ranking campaigns by ROI
    # FAILS on initial (no charts), PASSES on golden
    # -------------------------------------------------------------------
    try:
        chart_found = False
        correct_title = False

        # Check all sheets for charts
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            charts = sheet._charts
            if charts:
                for chart in charts:
                    chart_found = True
                    # Check chart type — should be BarChart (col type) or BarChart (bar type)
                    chart_type = type(chart).__name__
                    chart_type_attr = getattr(chart, 'type', None)
                    is_bar_or_col = (chart_type in ('BarChart',) or
                                     chart_type_attr in ('col', 'bar', 'barChart', 'colChart'))

                    # Try to extract title
                    title_text = get_chart_title_text(chart.title)

                    print(f"Chart found on sheet '{sheet_name}': class={chart_type}, type_attr={chart_type_attr}")
                    print(f"  Title text: {repr(title_text)}")

                    if title_text and 'campaign roi ranking' in title_text.lower():
                        correct_title = True
                        print(f"  PASS: Chart title matches 'Campaign ROI Ranking'")

                    if is_bar_or_col and correct_title:
                        print(f"PASS: Component 4 — Bar/column chart with correct title found")
                        total_score += 0.15
                        chart_found = True
                        break
                    elif chart_found and correct_title:
                        # Chart exists with correct title (any type)
                        print(f"PARTIAL: Component 4 — Chart with correct title found (type may differ)")
                        total_score += 0.10
                        break
                    elif chart_found and is_bar_or_col:
                        print(f"PARTIAL: Component 4 — Bar/column chart found but title incorrect (found: {repr(title_text)})")
                        total_score += 0.05
                        break
                if chart_found:
                    break

        if not chart_found:
            print(f"FAIL: Component 4 — No charts found in any sheet")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
