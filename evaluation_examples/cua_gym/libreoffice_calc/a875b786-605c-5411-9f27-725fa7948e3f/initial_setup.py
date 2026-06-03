"""
Initial Setup: Marketing Campaign ROI Dashboard (pre-task state)
Task ID: calc_sales_marketing_roi_026
Domain: libreoffice_calc

Creates a spreadsheet with 20 marketing campaigns.
Column H (ROI %) is intentionally left empty -- agent must add it.
No conditional formatting or charts are present.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_marketing_roi_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Campaigns'

    # --- Headers ---
    headers = [
        'Campaign ID', 'Campaign Name', 'Channel',
        'Cost', 'Leads Generated', 'Deals Won', 'Revenue Generated', 'ROI %'
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=11)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    # --- Campaign data (realistic marketing data) ---
    # 20 campaigns with varied channels, costs, leads, deals, revenue
    # ROI column (H) is intentionally left empty
    campaigns = [
        ('CMP-001', 'Spring Email Blast',         'Email',           12500,  340,  28,  185000),
        ('CMP-002', 'Google Search Q1',           'Paid Search',     85000,  920,  65,  520000),
        ('CMP-003', 'LinkedIn B2B Outreach',      'Social Media',    45000,  210,  18,   98000),
        ('CMP-004', 'Trade Show Chicago',         'Events',         120000,  580,  42,  310000),
        ('CMP-005', 'Facebook Retargeting',       'Social Media',    18000,  760,  55,  142000),
        ('CMP-006', 'Webinar Series Q1',          'Content',          9500,  430,  38,  215000),
        ('CMP-007', 'Influencer Partnership',     'Social Media',    55000,  640,  22,   88000),
        ('CMP-008', 'Direct Mail Campaign',       'Direct Mail',     32000,  190,  11,   41000),
        ('CMP-009', 'YouTube Pre-roll Ads',       'Video',           27500,  890,  48,  198000),
        ('CMP-010', 'SEO Content Push',           'Organic Search',   8200,  560,  44,  285000),
        ('CMP-011', 'Cold Outbound SDR',          'Outbound',        41000,  310,  19,   72000),
        ('CMP-012', 'Partner Co-Marketing',       'Partnerships',    15000,  480,  36,  175000),
        ('CMP-013', 'Podcast Sponsorship',        'Audio',           22000,  200,   8,   28000),
        ('CMP-014', 'Industry Conference NYC',    'Events',         150000,  720,  58,  850000),
        ('CMP-015', 'Twitter/X Promoted',         'Social Media',     5000,  310,   6,   12000),
        ('CMP-016', 'Product Hunt Launch',        'Online',           3200,  850,  70,  420000),
        ('CMP-017', 'Remarketing Display Ads',    'Display',         11500,  670,  41,  130000),
        ('CMP-018', 'Newsletter Sponsorship',     'Email',            6800,  290,  20,   58000),
        ('CMP-019', 'Case Study Downloads',       'Content',          4100,  380,  32,  160000),
        ('CMP-020', 'Radio Spot Campaign',        'Audio',           68000,  150,   5,    0),
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency_fmt = '$#,##0'
    int_fmt = '#,##0'

    for r, row_data in enumerate(campaigns, 2):
        cmp_id, name, channel, cost, leads, deals, revenue = row_data

        # Campaign ID
        c = ws.cell(row=r, column=1, value=cmp_id)
        c.alignment = Alignment(horizontal='center')
        c.border = data_border

        # Campaign Name
        c = ws.cell(row=r, column=2, value=name)
        c.border = data_border

        # Channel
        c = ws.cell(row=r, column=3, value=channel)
        c.alignment = Alignment(horizontal='center')
        c.border = data_border

        # Cost
        c = ws.cell(row=r, column=4, value=cost)
        c.number_format = currency_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = data_border

        # Leads Generated
        c = ws.cell(row=r, column=5, value=leads)
        c.number_format = int_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = data_border

        # Deals Won
        c = ws.cell(row=r, column=6, value=deals)
        c.number_format = int_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = data_border

        # Revenue Generated
        c = ws.cell(row=r, column=7, value=revenue)
        c.number_format = currency_fmt
        c.alignment = Alignment(horizontal='right')
        c.border = data_border

        # ROI % — intentionally empty (agent must add this)
        c = ws.cell(row=r, column=8, value=None)
        c.border = data_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Campaigns')
    print(f'  Rows: 20 campaign rows (rows 2-21)')
    print(f'  Column H (ROI %): empty (intentional)')


create_initial()
