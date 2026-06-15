"""
Initial Setup: Employee performance dataset with department names and monthly scores
Task ID: osworld_calc_multi_chart_computed_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Performance ---
    ws = wb.active
    ws.title = "Performance"

    # Headers: Department, Jan, Feb, Mar, Apr, May, Jun
    headers = ["Department", "Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name="Calibri", size=11)
        cell.alignment = Alignment(horizontal="center")

    # Employee data: 6 employees across 3 departments (2 per dept)
    # Department, Jan, Feb, Mar, Apr, May, Jun
    data = [
        # Engineering
        ["Engineering", 82, 85, 79, 88, 91, 87],
        ["Engineering", 75, 78, 81, 83, 80, 85],
        # Marketing
        ["Marketing", 88, 84, 90, 87, 93, 91],
        ["Marketing", 79, 82, 85, 88, 86, 90],
        # Sales
        ["Sales", 91, 89, 93, 95, 92, 96],
        ["Sales", 84, 87, 85, 89, 91, 88],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths
    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
