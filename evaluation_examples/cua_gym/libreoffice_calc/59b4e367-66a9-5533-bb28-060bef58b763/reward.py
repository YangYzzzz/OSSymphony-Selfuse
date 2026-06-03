"""
Reward Script: Calculate department averages and create bar+line charts
Task ID: osworld_calc_multi_chart_computed_007
Domain: libreoffice_calc
Scoring:
  - Component 1: Department average rows added with AVERAGE formulas (0.4 pts)
  - Component 2: Bar chart exists titled 'Department Average Performance' (0.3 pts)
  - Component 3: Line chart exists titled 'Performance Trend Over Time' (0.3 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_007'


def extract_chart_title(title_obj):
    """Extract text from an openpyxl chart title object."""
    try:
        if title_obj is None:
            return None
        for p in title_obj.tx.rich.p:
            texts = [r.t for r in p.r if r.t]
            if texts:
                return ''.join(texts)
    except Exception:
        return None
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Calculate average score per department in new rows, then create:
      - A bar chart titled 'Department Average Performance'
      - A line chart titled 'Performance Trend Over Time'
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Performance' sheet must exist
    if 'Performance' not in wb.sheetnames:
        print("CRITICAL: 'Performance' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Performance']

    # -----------------------------------------------------------------------
    # Component 1: Department average rows exist with AVERAGE formulas (0.4 pts)
    # The task requires adding average rows for each of the 3 departments.
    # In the golden file:
    #   Row 4: 'Engineering Average' with =AVERAGE(B2:B3) .. =AVERAGE(G2:G3)
    #   Row 7: 'Marketing Average'   with =AVERAGE(B5:B6) .. =AVERAGE(G5:G6)
    #   Row 10: 'Sales Average'      with =AVERAGE(B8:B9) .. =AVERAGE(G8:G9)
    # We check for the presence of department average label rows that contain
    # AVERAGE formulas in the score columns (B-G).
    # -----------------------------------------------------------------------
    try:
        dept_average_rows_found = 0
        expected_labels = ['Engineering Average', 'Marketing Average', 'Sales Average']

        # Scan all rows for average label rows that include AVERAGE formulas
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
            label_cell = row[0]  # Column A
            if label_cell.value is None:
                continue
            label_str = str(label_cell.value).strip()

            # Check if any expected average label is present (case-insensitive)
            is_avg_label = any(exp.lower() in label_str.lower() for exp in expected_labels)
            if not is_avg_label:
                # Fallback: check if label contains "Average" and a dept name
                if 'average' not in label_str.lower():
                    continue

            # Check that at least one of cols B-G contains an AVERAGE formula
            has_avg_formula = False
            for cell in row[1:7]:  # columns B through G
                val = cell.value
                if val and isinstance(val, str) and 'AVERAGE' in val.upper():
                    has_avg_formula = True
                    break

            if has_avg_formula:
                dept_average_rows_found += 1
                print(f"  Found average row: '{label_str}' (row {label_cell.row})")

        # Require at least 3 department average rows to earn full credit
        if dept_average_rows_found >= 3:
            print(f"PASS: Component 1 — {dept_average_rows_found} department average rows with AVERAGE formulas found (0.4 pts)")
            total_score += 0.4
        elif dept_average_rows_found >= 1:
            # Partial credit: some averages added but not all 3
            partial = round(0.4 * dept_average_rows_found / 3, 4)
            print(f"PARTIAL: Component 1 — only {dept_average_rows_found}/3 department average rows found (+{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — no department average rows with AVERAGE formulas found (expected 3)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Bar chart exists and is titled 'Department Average Performance'
    # (0.3 pts)
    # -----------------------------------------------------------------------
    try:
        bar_chart_found = False
        bar_title_correct = False
        expected_bar_title = 'Department Average Performance'

        # Charts may be in any sheet
        all_sheets_charts = []
        for sname in wb.sheetnames:
            for chart in wb[sname]._charts:
                all_sheets_charts.append((sname, chart))

        for sname, chart in all_sheets_charts:
            chart_class = type(chart).__name__
            if chart_class == 'BarChart':
                bar_chart_found = True
                title_text = extract_chart_title(chart.title)
                print(f"  Bar chart found in sheet '{sname}', title={title_text!r}")
                if title_text and expected_bar_title.lower() in title_text.lower():
                    bar_title_correct = True

        if bar_chart_found and bar_title_correct:
            print(f"PASS: Component 2 — Bar chart found with correct title '{expected_bar_title}' (0.3 pts)")
            total_score += 0.3
        elif bar_chart_found:
            # Chart exists but title is wrong — partial credit
            print(f"PARTIAL: Component 2 — Bar chart exists but title is incorrect (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — no BarChart found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Line chart exists and is titled 'Performance Trend Over Time'
    # (0.3 pts)
    # -----------------------------------------------------------------------
    try:
        line_chart_found = False
        line_title_correct = False
        expected_line_title = 'Performance Trend Over Time'

        for sname, chart in all_sheets_charts:
            chart_class = type(chart).__name__
            if chart_class == 'LineChart':
                line_chart_found = True
                title_text = extract_chart_title(chart.title)
                print(f"  Line chart found in sheet '{sname}', title={title_text!r}")
                if title_text and expected_line_title.lower() in title_text.lower():
                    line_title_correct = True

        if line_chart_found and line_title_correct:
            print(f"PASS: Component 3 — Line chart found with correct title '{expected_line_title}' (0.3 pts)")
            total_score += 0.3
        elif line_chart_found:
            # Chart exists but title is wrong
            print(f"PARTIAL: Component 3 — Line chart exists but title is incorrect (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — no LineChart found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point — run against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
