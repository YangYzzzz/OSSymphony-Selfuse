"""
Reward Script: Event Planning Budget with variance analysis, formulas, conditional formatting, and chart
Task ID: calc_wf_037
Domain: libreoffice_calc
Scoring:
  Component 1: Variance formulas in column E (0.20)
  Component 2: % of Budget formulas in column F (0.15)
  Component 3: Running Total column H (0.15)
  Component 4: Forecast column I with IF logic (0.10)
  Component 5: Cumulative Variance column J (0.10)
  Component 6: Category subtotals with SUMIF (0.10)
  Component 7: Conditional formatting on variance column (0.10)
  Component 8: Chart present with correct type and series (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_037'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # The task should use 'Event Budget' sheet
    if 'Event Budget' not in wb.sheetnames:
        print("CRITICAL: 'Event Budget' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Event Budget']

    # Data rows are 2-21 (20 line items)
    DATA_START = 2
    DATA_END = 21

    # Component 1: Variance formulas in column E (0.20 points)
    # Task: Variance = Estimated - Actual => E = C - D
    # Initial has E2:E21 as None, golden has =C-D formulas
    try:
        variance_count = 0
        for row in range(DATA_START, DATA_END + 1):
            val = ws.cell(row=row, column=5).value  # col E
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Check for =C<row>-D<row> pattern
                expected = f"=C{row}-D{row}"
                if normalized == expected.upper():
                    variance_count += 1
                # Also accept other valid variance formulas referencing C and D
                elif "C" in normalized and "D" in normalized and normalized.startswith("="):
                    variance_count += 1
        if variance_count >= 18:  # at least 18 of 20 rows
            print(f"PASS: Component 1 -- Variance formulas found in {variance_count}/20 rows (0.20 pts)")
            total_score += 0.20
        elif variance_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 1 -- Variance formulas found in {variance_count}/20 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Variance formulas found in only {variance_count}/20 rows")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: % of Budget formulas in column F (0.15 points)
    # Formula: =D<row>/$C$23 (Actual / Total Budget)
    try:
        pct_count = 0
        for row in range(DATA_START, DATA_END + 1):
            val = ws.cell(row=row, column=6).value  # col F
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Should reference D and C23 (total budget)
                if normalized.startswith("=") and "D" in normalized and "C" in normalized and "23" in normalized:
                    pct_count += 1
                # Also accept =D/15000 style
                elif normalized.startswith("=") and "D" in normalized and "15000" in normalized:
                    pct_count += 1
        if pct_count >= 18:
            print(f"PASS: Component 2 -- % of Budget formulas found in {pct_count}/20 rows (0.15 pts)")
            total_score += 0.15
        elif pct_count >= 10:
            partial = 0.07
            print(f"PARTIAL: Component 2 -- % of Budget formulas found in {pct_count}/20 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- % of Budget formulas found in only {pct_count}/20 rows")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Running Total column H (0.15 points)
    # Formula: =SUM($D$2:D<row>) -- cumulative sum of Actual costs
    try:
        running_count = 0
        # Check if H1 header exists
        h1 = ws.cell(row=1, column=8).value
        header_ok = h1 is not None and "running" in str(h1).lower()

        for row in range(DATA_START, DATA_END + 1):
            val = ws.cell(row=row, column=8).value  # col H
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Should be a cumulative SUM referencing column D
                if normalized.startswith("=") and "SUM" in normalized and "D" in normalized:
                    running_count += 1
        if running_count >= 18 and header_ok:
            print(f"PASS: Component 3 -- Running Total formulas in {running_count}/20 rows, header='{h1}' (0.15 pts)")
            total_score += 0.15
        elif running_count >= 18:
            print(f"PARTIAL: Component 3 -- Running Total formulas OK but header missing/wrong (0.10 pts)")
            total_score += 0.10
        elif running_count >= 10:
            partial = 0.07
            print(f"PARTIAL: Component 3 -- Running Total formulas in {running_count}/20 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Running Total formulas found in only {running_count}/20 rows")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Forecast column I with IF logic (0.10 points)
    # Formula: =IF(G="N",C,D) -- if not paid, use estimated; if paid, use actual
    try:
        forecast_count = 0
        i1 = ws.cell(row=1, column=9).value
        header_ok = i1 is not None and "forecast" in str(i1).lower()

        for row in range(DATA_START, DATA_END + 1):
            val = ws.cell(row=row, column=9).value  # col I
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Should be an IF formula referencing G (Paid column)
                if normalized.startswith("=") and "IF" in normalized and "G" in normalized:
                    forecast_count += 1
        if forecast_count >= 18 and header_ok:
            print(f"PASS: Component 4 -- Forecast IF formulas in {forecast_count}/20 rows (0.10 pts)")
            total_score += 0.10
        elif forecast_count >= 18:
            print(f"PARTIAL: Component 4 -- Forecast IF formulas OK but header missing/wrong (0.07 pts)")
            total_score += 0.07
        elif forecast_count >= 10:
            partial = 0.05
            print(f"PARTIAL: Component 4 -- Forecast IF formulas in {forecast_count}/20 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- Forecast IF formulas found in only {forecast_count}/20 rows")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Cumulative Variance column J (0.10 points)
    # Formula: =SUM($E$2:E<row>) -- running sum of variance
    try:
        cumvar_count = 0
        j1 = ws.cell(row=1, column=10).value
        # Accept various header names for cumulative variance
        header_ok = j1 is not None and ("cumul" in str(j1).lower() or "variance" in str(j1).lower())

        for row in range(DATA_START, DATA_END + 1):
            val = ws.cell(row=row, column=10).value  # col J
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Should be a cumulative SUM referencing column E
                if normalized.startswith("=") and "SUM" in normalized and "E" in normalized:
                    cumvar_count += 1
        if cumvar_count >= 18 and header_ok:
            print(f"PASS: Component 5 -- Cumulative Variance formulas in {cumvar_count}/20 rows (0.10 pts)")
            total_score += 0.10
        elif cumvar_count >= 18:
            print(f"PARTIAL: Component 5 -- Cumulative Variance formulas OK but header missing/wrong (0.07 pts)")
            total_score += 0.07
        elif cumvar_count >= 10:
            partial = 0.05
            print(f"PARTIAL: Component 5 -- Cumulative Variance formulas in {cumvar_count}/20 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- Cumulative Variance formulas found in only {cumvar_count}/20 rows")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Category subtotals with SUMIF (0.10 points)
    # Should have SUMIF formulas for category subtotals somewhere in the sheet
    try:
        sumif_count = 0
        # Scan rows 22-50 for SUMIF formulas (subtotals section below data)
        for row in range(22, min(ws.max_row + 1, 51)):
            for col in range(1, min(ws.max_column + 1, 11)):
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, str):
                    if "SUMIF" in val.upper():
                        sumif_count += 1

        # Need at least 6 SUMIF formulas (est + actual subtotals for 6 categories = 12, or at least 6)
        if sumif_count >= 6:
            print(f"PASS: Component 6 -- Found {sumif_count} SUMIF formulas for category subtotals (0.10 pts)")
            total_score += 0.10
        elif sumif_count >= 3:
            partial = 0.05
            print(f"PARTIAL: Component 6 -- Found {sumif_count} SUMIF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 -- Found only {sumif_count} SUMIF formulas")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # Component 7: Conditional formatting on variance column (0.10 points)
    # Should highlight negative variance (over-budget) in red
    try:
        cf_found = False
        for cf in ws.conditional_formatting:
            range_str = str(cf)
            # Check if conditional formatting covers variance column E
            if "E" in range_str:
                for rule in cf.rules:
                    # Should be a cellIs rule checking for < 0 (negative variance = over budget)
                    if rule.type == 'cellIs' and rule.operator in ('lessThan', 'lessThanOrEqual'):
                        cf_found = True
                        break
                    # Also accept formula-based rules
                    if rule.type == 'expression' or rule.type == 'cellIs':
                        cf_found = True
                        break
            if cf_found:
                break

        if cf_found:
            print(f"PASS: Component 7 -- Conditional formatting found on variance column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 -- No conditional formatting detected on variance column E")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    # Component 8: Chart present with correct type and series (0.10 points)
    # Task requires: clustered bars (est vs actual) and line overlay for cumulative variance
    # At minimum, check for presence of a chart with bar/column type and >=2 series
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            has_bar_type = chart.type in ('col', 'bar')
            has_multiple_series = len(chart.series) >= 2

            if has_bar_type and has_multiple_series:
                print(f"PASS: Component 8 -- Chart found: type={chart.type}, series={len(chart.series)} (0.10 pts)")
                total_score += 0.10
            elif has_bar_type:
                print(f"PARTIAL: Component 8 -- Chart type OK but only {len(chart.series)} series (0.05 pts)")
                total_score += 0.05
            else:
                print(f"PARTIAL: Component 8 -- Chart exists but type={chart.type}, expected col/bar (0.03 pts)")
                total_score += 0.03
        else:
            # Check all sheets for charts
            found_any = False
            for sn in wb.sheetnames:
                if len(wb[sn]._charts) > 0:
                    found_any = True
                    chart = wb[sn]._charts[0]
                    print(f"PARTIAL: Component 8 -- Chart found on sheet '{sn}', type={chart.type} (0.05 pts)")
                    total_score += 0.05
                    break
            if not found_any:
                print(f"FAIL: Component 8 -- No chart found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 8 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
