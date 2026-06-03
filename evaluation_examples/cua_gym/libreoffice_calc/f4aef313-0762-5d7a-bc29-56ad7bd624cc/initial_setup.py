"""
Initial Setup: Event Planning Budget Spreadsheet
Task ID: calc_wf_037
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Event Budget'

    # --- Headers ---
    headers = ['Category', 'Item', 'Estimated', 'Actual', 'Variance', '% of Budget', 'Paid']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data: 20 line items across 6 categories ---
    # Total budget = $15,000
    # Columns: Category, Item, Estimated, Actual, Variance(empty), % of Budget(empty), Paid
    data = [
        # Venue (3 items)
        ['Venue', 'Banquet Hall Rental', 3000, 3200, None, None, 'Y'],
        ['Venue', 'Setup & Cleanup Fee', 500, 450, None, None, 'Y'],
        ['Venue', 'Parking Lot Reservation', 300, 300, None, None, 'Y'],
        # Catering (4 items)
        ['Catering', 'Dinner Buffet (80 guests)', 2400, 2650, None, None, 'Y'],
        ['Catering', 'Appetizer Platters', 600, 580, None, None, 'Y'],
        ['Catering', 'Dessert Bar', 450, 475, None, None, 'Y'],
        ['Catering', 'Beverage Package', 800, 820, None, None, 'N'],
        # Decor (4 items)
        ['Decor', 'Floral Arrangements', 900, 1050, None, None, 'Y'],
        ['Decor', 'Table Linens & Chair Covers', 400, 380, None, None, 'Y'],
        ['Decor', 'Lighting & Draping', 600, 620, None, None, 'N'],
        ['Decor', 'Centerpieces', 350, 340, None, None, 'Y'],
        # Entertainment (3 items)
        ['Entertainment', 'Live Band (4 hours)', 1500, 1500, None, None, 'Y'],
        ['Entertainment', 'Sound System Rental', 400, 425, None, None, 'Y'],
        ['Entertainment', 'MC / Host', 300, 300, None, None, 'N'],
        # Photography (3 items)
        ['Photography', 'Lead Photographer (6 hrs)', 1200, 1200, None, None, 'Y'],
        ['Photography', 'Videographer', 800, 850, None, None, 'N'],
        ['Photography', 'Photo Booth Rental', 350, 375, None, None, 'Y'],
        # Miscellaneous (3 items)
        ['Miscellaneous', 'Invitations & Stationery', 250, 230, None, None, 'Y'],
        ['Miscellaneous', 'Transportation (Shuttle)', 500, 520, None, None, 'N'],
        ['Miscellaneous', 'Contingency / Misc Items', 400, 350, None, None, 'N'],
    ]

    data_font = Font(name='Calibri', size=11)
    currency_fmt = '$#,##0.00'

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c in (3, 4):  # Estimated, Actual
                cell.number_format = currency_fmt
            if c == 7:  # Paid column center
                cell.alignment = Alignment(horizontal='center')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 10

    # --- Row for total budget label ---
    total_row = 23
    ws.cell(row=total_row, column=1, value='Total Budget:').font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=total_row, column=3, value=15000).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=total_row, column=3).number_format = currency_fmt
    ws.cell(row=total_row, column=3).border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
