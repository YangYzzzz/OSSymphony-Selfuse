"""
FINAL REWARD SCRIPT - SUCCESS
Task: I want to freeze the range A1:F2 to keep my category headers and sub-categories always visible.
Generated: 2025-11-24 07:32:40
Status: success
Model: o3
Total Steps: 3
"""

import os
import sys
import openpyxl
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string, get_column_letter

"""
Reward Script: Verify that the user froze the range A1:F2
---------------------------------------------------------
Task requirement (LibreOffice Calc / Excel context):
    "I want to freeze the range A1:F2 to keep my category headers and
     sub-categories always visible."

Expected technical implementation:
    The worksheet’s freeze-pane should be positioned at cell G3 (column 7, row 3),
    so that all rows above (1-2) and columns to the left (A-F) remain visible
    while scrolling.

Verification strategy:
1. Locate the user-created .xlsx file (ignore any files containing the word
   "golden"; fall back gracefully if no alternative is found).
2. Load the workbook with openpyxl.
3. Inspect the active worksheet’s freeze_panes attribute.
4. Award partial credit (0.3) if *any* freeze pane is set.
5. Award full credit (1.0) only if the freeze pane is exactly at column 7 / row 3
   (cell G3).
6. Print detailed diagnostics for transparency.
7. Output the final reward as "REWARD: <score>".

The script returns a progressive score (0.0-1.0) and prints observable
verification steps, fully complying with anti-hacking rules.
"""


def find_target_file() -> str | None:
    """Locate the spreadsheet likely produced by the user.

    Search /home/user for .xlsx files, excluding those whose names contain
    "golden" (reference files). If multiple remain, prefer ones whose names
    suggest relevance (keywords), otherwise pick the first.
    """
    user_dir = "/home/user"
    if not os.path.isdir(user_dir):
        return None

    xlsx_files = [f for f in os.listdir(user_dir) if f.lower().endswith(".xlsx")]
    # Exclude golden/reference files
    candidates = [f for f in xlsx_files if "golden" not in f.lower()]

    if not candidates:
        # If the user overwrote the golden file, fall back to any .xlsx file
        candidates = xlsx_files
    if not candidates:
        return None

    if len(candidates) == 1:
        return os.path.join(user_dir, candidates[0])

    # Prefer filenames containing task-related keywords
    KEYWORDS = ("freeze", "category", "visible")
    for cand in candidates:
        lower = cand.lower()
        if any(k in lower for k in KEYWORDS):
            return os.path.join(user_dir, cand)

    # Still ambiguous – return the first candidate deterministically
    return os.path.join(user_dir, sorted(candidates)[0])


def verify_freeze_panes(file_path: str) -> float:
    print(f"Verifying freeze panes for file: {file_path}")
    max_score = 1.0
    score = 0.0

    # ---------- Load workbook ----------
    try:
        wb = openpyxl.load_workbook(file_path)
        print("✓ Workbook loaded successfully")
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active  # assume the user worked on the active sheet
    print(f"Active sheet detected: '{ws.title}'")

    # ---------- Check freeze_panes attribute ----------
    fp = ws.freeze_panes
    if fp is None:
        print("✗ No freeze panes set – headers will not stay visible")
        print("REWARD: 0.0")
        return 0.0

    print(f"✓ freeze_panes property found: {fp}")
    score += 0.3  # partial credit for attempting to freeze panes

    # ---------- Parse coordinate ----------
    try:
        if isinstance(fp, str):
            col_letters, row_num = coordinate_from_string(fp)
            col_idx = column_index_from_string(col_letters)
        else:  # fp is a Cell object
            col_idx = fp.column
            row_num = fp.row
    except Exception as e:
        print(f"✗ Could not parse freeze coordinate: {e}")
        print(f"REWARD: {score}")
        return score  # return partial score only

    print(f"Parsed freeze coordinate → Column: {col_idx}, Row: {row_num}")

    # Expected position for keeping A1:F2 visible is G3 (col 7, row 3)
    if col_idx == 7 and row_num == 3:
        print("✓ Freeze panes correctly set to G3 – Range A1:F2 will remain visible")
        score = max_score
    else:
        expected = "G3"
        actual = f"{get_column_letter(col_idx)}{row_num}"
        print(f"✗ Incorrect freeze location. Expected {expected}, found {actual}")
        # keep partial score (0.3)

    final_score = min(score, max_score)
    print(f"REWARD: {final_score}")
    return final_score


if __name__ == "__main__":
    # Allow CLI override: `python verify.py my_file.xlsx`
    target_file = sys.argv[1] if len(sys.argv) > 1 else find_target_file()

    if not target_file or not os.path.exists(target_file):
        print("✗ Could not locate the target .xlsx file for verification")
        print("REWARD: 0.0")
        sys.exit()

    verify_freeze_panes(target_file)

