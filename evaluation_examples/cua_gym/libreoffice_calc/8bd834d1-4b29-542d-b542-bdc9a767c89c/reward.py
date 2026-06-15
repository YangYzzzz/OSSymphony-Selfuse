"""
Reward Script: Visit faculty pages to fill email and office hours in spreadsheet
Task ID: osworld_multi_apps_web_prof_email_013
Domain: libreoffice_calc
Scoring:
  - Component 1: All 6 email cells are populated with valid email addresses (0.5 pts)
  - Component 2: All 6 office hours cells are populated (non-empty) (0.3 pts)
  - Component 3: Emails match expected institutional format for at least 5/6 rows (0.2 pts)
Total: 1.0
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_013'

# Expected golden-state email addresses (from the golden_patch)
EXPECTED_EMAILS = {
    'Dan Boneh': 'dabo@cs.stanford.edu',
    'Yael Tauman Kalai': 'yael@csail.mit.edu',
    'Vitaly Shmatikov': 'shmat@cs.cornell.edu',
    'Hovav Shacham': 'hovav@cs.utexas.edu',
    'Stefan Savage': 'savage@cs.ucsd.edu',
    'Nikita Borisov': 'nikita@illinois.edu',
}

EMAIL_PATTERN = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')


def is_valid_email(value):
    """Return True if value is a non-empty string that looks like an email."""
    if not value or not isinstance(value, str):
        return False
    return bool(EMAIL_PATTERN.match(value.strip()))


def is_populated(value):
    """Return True if value is a non-empty, non-None string."""
    if value is None:
        return False
    val = str(value).strip()
    return len(val) > 0


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the sheet
    try:
        ws = wb['Security_Researchers']
    except KeyError:
        # Fallback to active sheet
        ws = wb.active

    # Verify header structure (precondition gate)
    try:
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        expected_headers = ['Name', 'University', 'Faculty Page URL', 'Email', 'Office Hours']
        if headers != expected_headers:
            print(f"CRITICAL: Unexpected headers: {headers}")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"CRITICAL: Cannot read headers: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Collect data rows (rows 2-7, 6 researchers)
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=7, min_col=1, max_col=5):
        name = row[0].value
        email = row[3].value      # column D = index 3
        office_hours = row[4].value  # column E = index 4
        data_rows.append((name, email, office_hours))

    if len(data_rows) == 0:
        print("CRITICAL: No data rows found in spreadsheet.")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All 6 email cells populated with valid email addresses (0.5 pts)
    try:
        emails_valid = [(name, email) for name, email, _ in data_rows if is_valid_email(email)]
        emails_count = len(emails_valid)
        total_rows = len(data_rows)

        if emails_count == total_rows:
            print(f"PASS: Component 1 — All {total_rows} email cells contain valid email addresses (0.5 pts)")
            total_score += 0.5
        elif emails_count > 0:
            partial = round(0.5 * emails_count / total_rows, 3)
            print(f"PARTIAL: Component 1 — {emails_count}/{total_rows} emails are valid; partial credit {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No valid email addresses found in Email column; emails_count=0")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 6 office hours cells populated (non-empty) (0.3 pts)
    try:
        oh_populated = [(name, oh) for name, _, oh in data_rows if is_populated(oh)]
        oh_count = len(oh_populated)

        if oh_count == total_rows:
            print(f"PASS: Component 2 — All {total_rows} office hours cells are populated (0.3 pts)")
            total_score += 0.3
        elif oh_count > 0:
            partial = round(0.3 * oh_count / total_rows, 3)
            print(f"PARTIAL: Component 2 — {oh_count}/{total_rows} office hours cells populated; partial credit {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No office hours values found in Office Hours column; oh_count=0")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: At least 5/6 emails match expected institutional values (0.2 pts)
    # This verifies that the correct researcher emails were actually found (not arbitrary values)
    try:
        matches = 0
        for name, email, _ in data_rows:
            if name in EXPECTED_EMAILS:
                expected = EXPECTED_EMAILS[name]
                actual = str(email).strip().lower() if email else ''
                if actual == expected.lower():
                    matches += 1
                else:
                    print(f"  NOTE: Component 3 — {name}: expected '{expected}', found '{email}'")

        threshold = 5  # at least 5 of 6 must match
        if matches >= threshold:
            print(f"PASS: Component 3 — {matches}/{len(data_rows)} emails match expected institutional addresses (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {matches}/{len(data_rows)} emails match expected values (need >= {threshold})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 3), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
