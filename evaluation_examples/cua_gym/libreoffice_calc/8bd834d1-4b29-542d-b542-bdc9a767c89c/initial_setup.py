"""
Initial Setup: Security researchers spreadsheet with Chrome open for web browsing
Task ID: osworld_multi_apps_web_prof_email_013
Domain: libreoffice_calc (multi-app with Chrome)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Security_Researchers ---
    ws = wb.active
    ws.title = 'Security_Researchers'

    # Headers
    headers = ['Name', 'University', 'Faculty Page URL', 'Email', 'Office Hours']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 55
    ws.column_dimensions['D'].width = 32
    ws.column_dimensions['E'].width = 28

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Security researcher data — Email and Office Hours are intentionally blank
    # These are real faculty pages for well-known security researchers
    researchers = [
        [
            'Dan Boneh',
            'Stanford University',
            'https://crypto.stanford.edu/~dabo/',
            '',  # Email — to be filled by agent
            '',  # Office Hours — to be filled by agent
        ],
        [
            'Yael Tauman Kalai',
            'MIT',
            'https://people.csail.mit.edu/yael/',
            '',
            '',
        ],
        [
            'Vitaly Shmatikov',
            'Cornell University',
            'https://www.cs.cornell.edu/~shmat/',
            '',
            '',
        ],
        [
            'Hovav Shacham',
            'University of Texas at Austin',
            'https://cs.utexas.edu/~hovav/',
            '',
            '',
        ],
        [
            'Stefan Savage',
            'UC San Diego',
            'https://cseweb.ucsd.edu/~savage/',
            '',
            '',
        ],
        [
            'Nikita Borisov',
            'University of Illinois Urbana-Champaign',
            'https://hatswitch.org/~nikita/',
            '',
            '',
        ],
    ]

    # Data font and alignment
    data_font = Font(size=11)
    data_align = Alignment(vertical='center')
    url_align = Alignment(vertical='center', horizontal='left', wrap_text=False)

    for row_idx, row_data in enumerate(researchers, 2):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            if col_idx == 3:
                cell.alignment = url_align
                cell.font = Font(size=11, color='FF0563C1', underline='single')
            else:
                cell.alignment = data_align

        # Alternating row shading
        if row_idx % 2 == 0:
            row_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
            for col_idx in range(1, 6):
                ws.cell(row=row_idx, column=col_idx).fill = row_fill

    # Row height
    ws.row_dimensions[1].height = 22
    for r in range(2, 8):
        ws.row_dimensions[r].height = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open LibreOffice Calc with the spreadsheet, then Chrome
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui('google-chrome', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
