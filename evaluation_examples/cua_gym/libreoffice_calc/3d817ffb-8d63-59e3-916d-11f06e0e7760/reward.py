"""
Reward Script: Format trial balance with currency, totals, difference check, conditional formatting, and row grouping
Task ID: calc_fin_audit_trail_041
Domain: libreoffice_calc
Scoring:
  Component 1: Currency format on D2:E60          (0.20 pts)
  Component 2: TOTALS row 61 with SUM formulas    (0.25 pts)
  Component 3: Difference row 62 with formula     (0.20 pts)
  Component 4: Conditional formatting on B62      (0.15 pts)
  Component 5: Row grouping rows 2-60             (0.10 pts)
  Component 6: Freeze panes at A2                 (0.05 pts)
  Component 7: Row 1 bold                         (0.05 pts)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_audit_trail_041'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheet exists
    if 'TrialBalance' not in wb.sheetnames:
        print("CRITICAL: 'TrialBalance' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TrialBalance']

    # Component 1: Currency format on D2:E60 (0.20 points)
    # Task requires D2:E60 to be currency formatted (initially 'General')
    try:
        currency_formats = ['$#,##0.00', '#,##0.00', '[$USD]#,##0.00']
        currency_count = 0
        total_cells = 0
        for row in range(2, 61):
            for col in [4, 5]:  # D and E
                cell = ws.cell(row=row, column=col)
                total_cells += 1
                nf = cell.number_format
                # Check for currency-like format (contains $ or comma formatting)
                if nf and ('$' in nf or nf in currency_formats or '#,##0' in nf):
                    currency_count += 1
        if total_cells > 0 and currency_count >= total_cells * 0.9:
            print(f"PASS: Component 1 — Currency format on D2:E60 ({currency_count}/{total_cells} cells) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Expected currency format on D2:E60, found {currency_count}/{total_cells} correctly formatted")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: TOTALS row 61 with SUM formulas and bold (0.25 points)
    # Task requires A61='TOTALS', D61=SUM(D2:D60), E61=SUM(E2:E60), all bold
    try:
        a61_val = ws['A61'].value
        d61_val = ws['D61'].value
        e61_val = ws['E61'].value
        a61_bold = ws['A61'].font.bold
        d61_bold = ws['D61'].font.bold
        e61_bold = ws['E61'].font.bold

        totals_label_ok = (a61_val is not None and str(a61_val).strip().upper() == 'TOTALS')
        d61_formula_ok = (d61_val is not None and isinstance(d61_val, str) and
                          'SUM' in d61_val.upper() and 'D2' in d61_val.upper() and 'D60' in d61_val.upper())
        e61_formula_ok = (e61_val is not None and isinstance(e61_val, str) and
                          'SUM' in e61_val.upper() and 'E2' in e61_val.upper() and 'E60' in e61_val.upper())
        bold_ok = (a61_bold == True and d61_bold == True and e61_bold == True)

        checks_passed = sum([totals_label_ok, d61_formula_ok, e61_formula_ok, bold_ok])
        if checks_passed == 4:
            print(f"PASS: Component 2 — TOTALS row 61 fully correct (label, SUM formulas, bold) (0.25 pts)")
            total_score += 0.25
        elif checks_passed >= 2:
            print(f"PARTIAL: Component 2 — TOTALS row 61 partially correct ({checks_passed}/4 checks passed)")
            print(f"  label={totals_label_ok} (A61={repr(a61_val)}), D61_formula={d61_formula_ok} ({repr(d61_val)}), E61_formula={e61_formula_ok} ({repr(e61_val)}), bold={bold_ok}")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — TOTALS row 61 missing or incorrect ({checks_passed}/4 checks passed)")
            print(f"  label={totals_label_ok} (A61={repr(a61_val)}), D61_formula={d61_formula_ok} ({repr(d61_val)}), E61_formula={e61_formula_ok} ({repr(e61_val)}), bold={bold_ok}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Difference row 62 with formula and bold (0.20 points)
    # Task requires A62='Difference (should be 0)', B62=D61-E61, both bold
    try:
        a62_val = ws['A62'].value
        b62_val = ws['B62'].value
        a62_bold = ws['A62'].font.bold
        b62_bold = ws['B62'].font.bold

        diff_label_ok = (a62_val is not None and 'difference' in str(a62_val).lower())
        b62_formula_ok = (b62_val is not None and isinstance(b62_val, str) and
                          'D61' in b62_val.upper() and 'E61' in b62_val.upper())
        bold_ok = (a62_bold == True and b62_bold == True)

        checks_passed = sum([diff_label_ok, b62_formula_ok, bold_ok])
        if checks_passed == 3:
            print(f"PASS: Component 3 — Difference row 62 fully correct (label, formula, bold) (0.20 pts)")
            total_score += 0.20
        elif checks_passed >= 2:
            print(f"PARTIAL: Component 3 — Difference row 62 partially correct ({checks_passed}/3 checks)")
            print(f"  label={diff_label_ok} (A62={repr(a62_val)}), formula={b62_formula_ok} (B62={repr(b62_val)}), bold={bold_ok}")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Difference row 62 missing or incorrect ({checks_passed}/3 checks)")
            print(f"  label={diff_label_ok} (A62={repr(a62_val)}), formula={b62_formula_ok} (B62={repr(b62_val)}), bold={bold_ok}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on B62 (0.15 points)
    # Task requires: conditional formatting on B62 with expression B62<>0 => red background
    try:
        cf_list = list(ws.conditional_formatting)
        b62_cf_found = False
        b62_cf_has_formula = False
        b62_cf_has_red_fill = False

        for cf in cf_list:
            cf_str = str(cf)
            # Check if this CF applies to B62
            if 'B62' in cf_str:
                b62_cf_found = True
                for rule in cf.rules:
                    formula = getattr(rule, 'formula', None)
                    if formula and any('B62' in f.upper() for f in formula):
                        b62_cf_has_formula = True
                    # Check for red fill
                    try:
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fill = rule.dxf.fill
                            if fill.fgColor and fill.fgColor.rgb:
                                rgb = fill.fgColor.rgb.upper()
                                # Red fill: FFFF0000 or variations
                                if 'FF0000' in rgb or rgb == 'FFFF0000':
                                    b62_cf_has_red_fill = True
                    except Exception:
                        pass

        checks_passed = sum([b62_cf_found, b62_cf_has_formula, b62_cf_has_red_fill])
        if checks_passed == 3:
            print(f"PASS: Component 4 — Conditional formatting on B62 with expression and red fill (0.15 pts)")
            total_score += 0.15
        elif checks_passed >= 1:
            print(f"PARTIAL: Component 4 — CF partially correct ({checks_passed}/3 checks): found={b62_cf_found}, formula={b62_cf_has_formula}, red_fill={b62_cf_has_red_fill}")
            total_score += 0.07
        else:
            print(f"FAIL: Component 4 — No conditional formatting on B62 found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Row grouping (rows 2-60 have outlineLevel=1) (0.10 points)
    # Task requires groups for Assets(2-15), Liabilities(16-25), Equity(26-30), Revenue(31-40), Expenses(41-60)
    try:
        grouped_rows = []
        for row_idx, rd in ws.row_dimensions.items():
            if rd.outlineLevel >= 1:
                grouped_rows.append(row_idx)

        # Expect rows 2-60 to be grouped (59 rows)
        expected_grouped = set(range(2, 61))
        actual_grouped = set(grouped_rows)
        overlap = len(expected_grouped.intersection(actual_grouped))
        total_expected = len(expected_grouped)

        if overlap >= total_expected * 0.9:
            print(f"PASS: Component 5 — Row grouping correct ({overlap}/{total_expected} rows grouped at outlineLevel>=1) (0.10 pts)")
            total_score += 0.10
        elif overlap >= total_expected * 0.5:
            print(f"PARTIAL: Component 5 — Row grouping partially correct ({overlap}/{total_expected} rows)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Row grouping insufficient ({overlap}/{total_expected} rows grouped)")
            print(f"  Grouped rows found: {sorted(grouped_rows)[:10]}...")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Freeze panes at A2 (0.05 points)
    # Task requires Row 1 frozen so header stays visible
    try:
        freeze = ws.freeze_panes
        if freeze == "A2":
            print(f"PASS: Component 6 — Freeze panes at A2 (header row frozen) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Expected freeze_panes='A2', found: {repr(freeze)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Row 1 bold (0.05 points)
    # Task requires header row to be bold
    try:
        row1_bold_cells = 0
        for col in range(1, 6):  # A-E
            if ws.cell(row=1, column=col).font.bold:
                row1_bold_cells += 1
        if row1_bold_cells >= 3:
            print(f"PASS: Component 7 — Row 1 headers bold ({row1_bold_cells}/5 cells bold) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — Row 1 headers not bold ({row1_bold_cells}/5 cells bold)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
