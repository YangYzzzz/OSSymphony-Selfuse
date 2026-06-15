"""
Initial Setup: Financial Audit Trial Balance
Task ID: calc_fin_audit_trail_041
Domain: libreoffice_calc

Creates a trial balance spreadsheet with chart of accounts entries
grouped by type (Assets, Liabilities, Equity, Revenue, Expenses).
NO formatting, no totals, no grouping, no frozen panes applied.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_audit_trail_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TrialBalance'

    # --- Row 1: Headers (NOT bold, NOT frozen) ---
    headers = ['Account#', 'Account Name', 'Type', 'Debit', 'Credit']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Data: 59 entries across 5 account types ---
    # Format: [Account#, Account Name, Type, Debit, Credit]
    # Assets: rows 2-15 (14 entries)
    assets_data = [
        ['1010', 'Cash and Cash Equivalents',       'Assets',  125430.00,      0.00],
        ['1020', 'Petty Cash',                       'Assets',    2500.00,      0.00],
        ['1110', 'Accounts Receivable - Trade',      'Assets',  348920.50,      0.00],
        ['1115', 'Allowance for Doubtful Accounts',  'Assets',       0.00,  12500.00],
        ['1120', 'Notes Receivable',                 'Assets',   75000.00,      0.00],
        ['1210', 'Inventory - Finished Goods',       'Assets',  215670.80,      0.00],
        ['1215', 'Inventory - Raw Materials',        'Assets',   89340.25,      0.00],
        ['1220', 'Inventory - Work in Progress',     'Assets',   43210.00,      0.00],
        ['1310', 'Prepaid Insurance',                'Assets',    8750.00,      0.00],
        ['1320', 'Prepaid Rent',                     'Assets',   12000.00,      0.00],
        ['1410', 'Property, Plant & Equipment',      'Assets', 1250000.00,      0.00],
        ['1415', 'Accumulated Depreciation - PPE',   'Assets',       0.00, 425000.00],
        ['1420', 'Intangible Assets - Patents',      'Assets',   95000.00,      0.00],
        ['1425', 'Accumulated Amortization',         'Assets',       0.00,  18000.00],
    ]

    # Liabilities: rows 16-25 (10 entries)
    liabilities_data = [
        ['2010', 'Accounts Payable - Trade',         'Liabilities',      0.00, 187650.30],
        ['2020', 'Accrued Salaries Payable',         'Liabilities',      0.00,  45320.00],
        ['2030', 'Accrued Interest Payable',         'Liabilities',      0.00,   8250.75],
        ['2040', 'Income Taxes Payable',             'Liabilities',      0.00,  32400.00],
        ['2050', 'Deferred Revenue',                 'Liabilities',      0.00,  28700.00],
        ['2060', 'Short-term Notes Payable',         'Liabilities',      0.00,  50000.00],
        ['2070', 'Current Portion of Long-term Debt','Liabilities',      0.00,  75000.00],
        ['2110', 'Long-term Notes Payable',          'Liabilities',      0.00, 250000.00],
        ['2120', 'Mortgage Payable',                 'Liabilities',      0.00, 480000.00],
        ['2130', 'Deferred Tax Liability',           'Liabilities',      0.00,  22500.00],
    ]

    # Equity: rows 26-30 (5 entries)
    equity_data = [
        ['3010', 'Common Stock - Par Value',         'Equity',           0.00, 100000.00],
        ['3020', 'Additional Paid-in Capital',       'Equity',           0.00, 350000.00],
        ['3030', 'Retained Earnings - Prior Year',   'Equity',           0.00, 198450.00],
        ['3040', 'Current Year Net Income',          'Equity',           0.00,  82340.00],
        ['3050', 'Treasury Stock',                   'Equity',       25000.00,      0.00],
    ]

    # Revenue: rows 31-40 (10 entries)
    revenue_data = [
        ['4010', 'Product Sales Revenue',            'Revenue',          0.00, 875420.00],
        ['4020', 'Service Revenue',                  'Revenue',          0.00, 234560.00],
        ['4030', 'Subscription Revenue',             'Revenue',          0.00,  98320.00],
        ['4040', 'Licensing Revenue',                'Revenue',          0.00,  45600.00],
        ['4050', 'Rental Income',                    'Revenue',          0.00,  36000.00],
        ['4060', 'Interest Income',                  'Revenue',          0.00,   7840.50],
        ['4070', 'Dividend Income',                  'Revenue',          0.00,   3200.00],
        ['4080', 'Gain on Sale of Assets',           'Revenue',          0.00,  12500.00],
        ['4090', 'Foreign Exchange Gain',            'Revenue',          0.00,   5680.25],
        ['4095', 'Miscellaneous Revenue',            'Revenue',          0.00,   8750.00],
    ]

    # Expenses: rows 41-60 (20 entries)
    expenses_data = [
        ['5010', 'Cost of Goods Sold',               'Expenses',    524250.00,      0.00],
        ['5020', 'Direct Labor',                     'Expenses',    142800.00,      0.00],
        ['5030', 'Manufacturing Overhead',           'Expenses',     68430.00,      0.00],
        ['5110', 'Salaries & Wages - Admin',         'Expenses',    186500.00,      0.00],
        ['5120', 'Salaries & Wages - Sales',         'Expenses',    124300.00,      0.00],
        ['5130', 'Employee Benefits Expense',        'Expenses',     48920.00,      0.00],
        ['5140', 'Payroll Tax Expense',              'Expenses',     29140.00,      0.00],
        ['5210', 'Rent Expense - Office',            'Expenses',     36000.00,      0.00],
        ['5220', 'Utilities Expense',                'Expenses',     18450.00,      0.00],
        ['5230', 'Telephone & Internet',             'Expenses',      6780.00,      0.00],
        ['5240', 'Office Supplies Expense',          'Expenses',      4320.00,      0.00],
        ['5310', 'Depreciation Expense',             'Expenses',     85000.00,      0.00],
        ['5320', 'Amortization Expense',             'Expenses',      9500.00,      0.00],
        ['5410', 'Advertising & Marketing',          'Expenses',     52600.00,      0.00],
        ['5420', 'Travel & Entertainment',           'Expenses',     18900.00,      0.00],
        ['5430', 'Professional Services',            'Expenses',     34500.00,      0.00],
        ['5440', 'Insurance Expense',                'Expenses',     22800.00,      0.00],
        ['5510', 'Interest Expense',                 'Expenses',     28350.00,      0.00],
        ['5520', 'Bank Charges',                     'Expenses',      2150.00,      0.00],
        ['5530', 'Income Tax Expense',               'Expenses',     32400.00,      0.00],
    ]

    all_data = assets_data + liabilities_data + equity_data + revenue_data + expenses_data

    for r, row_data in enumerate(all_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: TrialBalance')
    print(f'  Rows: 1 header + 59 data rows (rows 2-60)')
    print(f'  Assets: rows 2-15 (14 entries)')
    print(f'  Liabilities: rows 16-25 (10 entries)')
    print(f'  Equity: rows 26-30 (5 entries)')
    print(f'  Revenue: rows 31-40 (10 entries)')
    print(f'  Expenses: rows 41-60 (20 entries)')
    print(f'  NO formatting, NO totals, NO grouping, NO frozen panes')


create_initial()
