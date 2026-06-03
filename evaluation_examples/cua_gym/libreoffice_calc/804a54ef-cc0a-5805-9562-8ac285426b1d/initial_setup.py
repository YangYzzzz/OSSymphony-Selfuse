"""
Initial Setup: Add Y-error bars to 'Experiment A' data series in an embedded chart
Task ID: calc_gg2_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.chart import LineChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Lab Results ---
    ws = wb.active
    ws.title = 'Lab Results'

    # Headers
    headers = ['Time Point', 'Experiment A', 'Experiment B', 'Experiment C']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data: 10 time points with realistic lab measurement values
    data = [
        [1,  23.4, 31.2, 45.8],
        [2,  27.1, 33.5, 48.3],
        [3,  30.8, 29.7, 52.1],
        [4,  35.2, 35.0, 49.6],
        [5,  38.9, 37.8, 55.2],
        [6,  42.3, 34.1, 58.7],
        [7,  45.7, 38.6, 53.4],
        [8,  49.1, 40.2, 61.0],
        [9,  52.6, 36.9, 57.8],
        [10, 56.0, 42.5, 64.3],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    # --- Create an embedded line chart (NO error bars) ---
    chart = LineChart()
    chart.title = "Lab Experiment Results"
    chart.y_axis.title = "Measurement Value"
    chart.x_axis.title = "Time Point"
    chart.width = 18
    chart.height = 12

    # Data series: columns B, C, D (rows 1-11, titles from data)
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=4, max_row=11)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=11)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Style the series
    chart.series[0].graphicalProperties.line.width = 25000  # Experiment A
    chart.series[1].graphicalProperties.line.width = 25000  # Experiment B
    chart.series[2].graphicalProperties.line.width = 25000  # Experiment C

    ws.add_chart(chart, "A14")

    # --- Sheet 2: Metadata ---
    ws2 = wb.create_sheet('Metadata')
    ws2['A1'] = 'Study Information'
    ws2['A2'] = 'Principal Investigator'
    ws2['B2'] = 'Dr. Elena Vasquez'
    ws2['A3'] = 'Lab'
    ws2['B3'] = 'Biomedical Engineering - Lab 204'
    ws2['A4'] = 'Date Range'
    ws2['B4'] = 'March 10-19, 2025'
    ws2['A5'] = 'Instrument'
    ws2['B5'] = 'Shimadzu UV-2600i Spectrophotometer'
    ws2['A6'] = 'Units'
    ws2['B6'] = 'Absorbance (AU x 10^3)'
    ws2['A7'] = 'Notes'
    ws2['B7'] = 'Experiment A has known measurement uncertainty of +/-5.0 AU across all time points'
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 45

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
