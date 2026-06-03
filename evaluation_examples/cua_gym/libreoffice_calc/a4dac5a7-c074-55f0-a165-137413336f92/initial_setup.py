"""
Initial Setup: Payment Confirmation Email Tracker - Thunderbird + LibreOffice Calc Task
Task ID: osworld_multi_apps_email_data_002
Domain: libreoffice_calc (multi-app: Thunderbird + LibreOffice Calc)

Creates:
  1. /home/user/finance_tracker.ods with headers in row 1 (Subject/Sender/Date),
     A2 is intentionally EMPTY (agent must copy from Thunderbird).
  2. Thunderbird profile with 3 emails in a 'Finance' folder.
     The first email (newest, shown first in default Thunderbird sort) has
     Subject: "Payment Confirmation - Invoice #2024-0847"
  3. Opens Thunderbird (to Finance folder) and finance_tracker.ods in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time
import glob as globmod
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_data_002'
OUTPUT = f'{WORKDIR}/finance_tracker.ods'

# The 3 Finance folder emails (newest first — first email in default Thunderbird view)
# GROUND TRUTH: The first email (index 0, newest date) subject is used in golden
FINANCE_EMAILS = [
    {
        "subject": "Payment Confirmation - Invoice #2024-0847",
        "from": "billing@supplierco.com",
        "date": "Mon, 03 Mar 2026 10:15:00 +0000",
        "body": (
            "Dear Finance Team,\n\n"
            "We are pleased to confirm that payment for Invoice #2024-0847 "
            "has been successfully processed.\n\n"
            "Invoice #2024-0847\n"
            "Amount: $3,450.00\n"
            "Payment Method: Bank Transfer\n"
            "Reference: TRX-20260303-0047\n\n"
            "Please keep this confirmation for your records.\n\n"
            "Best regards,\n"
            "SupplierCo Billing Department"
        ),
    },
    {
        "subject": "Invoice #2024-0831 - Payment Due Reminder",
        "from": "accounts@globaltech.net",
        "date": "Wed, 26 Feb 2026 14:00:00 +0000",
        "body": (
            "Hello,\n\n"
            "This is a friendly reminder that Invoice #2024-0831 is due "
            "for payment within 7 days.\n\n"
            "Invoice #2024-0831\n"
            "Service: Software License Renewal Q1 2026\n"
            "Amount Due: $1,200.00\n"
            "Due Date: March 5, 2026\n\n"
            "Please contact us if you have any questions.\n\n"
            "Regards,\n"
            "GlobalTech Accounts Receivable"
        ),
    },
    {
        "subject": "Receipt: Annual Subscription Renewal #SUB-2026-114",
        "from": "noreply@cloudservices.io",
        "date": "Fri, 21 Feb 2026 09:30:00 +0000",
        "body": (
            "Thank you for renewing your subscription.\n\n"
            "Subscription: CloudServices Pro Plan\n"
            "Reference: SUB-2026-114\n"
            "Period: Feb 21, 2026 - Feb 20, 2027\n"
            "Total Charged: $599.00\n"
            "Card: Visa ending in 4821\n\n"
            "Your subscription is now active.\n\n"
            "CloudServices Team"
        ),
    },
]

# Ground truth: the first email subject (newest, shown at top in Thunderbird)
FIRST_EMAIL_SUBJECT = FINANCE_EMAILS[0]["subject"]


def create_initial_ods():
    """Create finance_tracker.ods with headers only — A2 is EMPTY (task not done yet)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers in row 1 (as specified in task context)
    ws['A1'] = 'Subject'
    ws['B1'] = 'Sender'
    ws['C1'] = 'Date'

    # A2, B2, C2 are intentionally empty — agent must fill A2 from Thunderbird

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 20

    wb.save(OUTPUT)
    print(f'Initial ODS file created: {OUTPUT}')


def setup_thunderbird_finance_folder():
    """
    Inject 3 emails into Thunderbird's Finance local folder using mbox format.
    Creates the Finance folder under Local Folders.
    Emails are stored newest-first in mbox so Thunderbird default date-desc
    sort shows them in the same order.
    """
    # Thunderbird profile path on OSWorld VM
    tb_profile_base = os.path.expanduser('~/.thunderbird')

    # Find the default profile directory
    profile_dir = None
    profiles_ini = os.path.join(tb_profile_base, 'profiles.ini')
    if os.path.exists(profiles_ini):
        with open(profiles_ini, 'r') as f:
            content = f.read()
        # Parse all profiles, prefer the default-release one
        sections = {}
        current_section = None
        for line in content.split('\n'):
            line = line.strip()
            if line.startswith('[') and line.endswith(']'):
                current_section = line[1:-1]
                sections[current_section] = {}
            elif '=' in line and current_section:
                k, v = line.split('=', 1)
                sections[current_section][k.strip()] = v.strip()
        # Find profile with Default=1 or use the last Path found
        for sec_name, sec_vals in sections.items():
            if sec_name.startswith('Profile') and 'Path' in sec_vals:
                rel_path = sec_vals['Path']
                is_relative = sec_vals.get('IsRelative', '1') == '1'
                if is_relative:
                    candidate = os.path.join(tb_profile_base, rel_path)
                else:
                    candidate = rel_path
                if os.path.isdir(candidate):
                    profile_dir = candidate
                    if sec_vals.get('Default', '0') == '1':
                        break  # prefer the default profile

    if not profile_dir or not os.path.isdir(profile_dir):
        candidates = globmod.glob(os.path.join(tb_profile_base, '*.default*'))
        candidates += globmod.glob(os.path.join(tb_profile_base, '*.default'))
        if candidates:
            profile_dir = sorted(candidates)[-1]
        else:
            profile_dir = os.path.join(tb_profile_base, 'default.default')
            os.makedirs(profile_dir, exist_ok=True)
            with open(profiles_ini, 'w') as f:
                f.write('[General]\nStartWithLastProfile=1\n\n'
                        '[Profile0]\nName=default\nIsRelative=1\n'
                        'Path=default.default\nDefault=1\n')

    print(f'Using Thunderbird profile: {profile_dir}')

    # Create the Finance folder under Local Folders
    mail_dir = os.path.join(profile_dir, 'Mail', 'Local Folders')
    os.makedirs(mail_dir, exist_ok=True)

    finance_path = os.path.join(mail_dir, 'Finance')

    # Build mbox content — emails stored newest-first
    mbox_lines = []
    for idx, email in enumerate(FINANCE_EMAILS):
        from_addr = email['from'].replace(' ', '_')
        mbox_lines.append(f"From {from_addr} {email['date']}\n")
        mbox_lines.append(f"From: {email['from']}\n")
        mbox_lines.append(f"To: finance@company.com\n")
        mbox_lines.append(f"Subject: {email['subject']}\n")
        mbox_lines.append(f"Date: {email['date']}\n")
        mbox_lines.append(f"Message-ID: <{TASK_ID}-{idx}@localhost>\n")
        mbox_lines.append(f"MIME-Version: 1.0\n")
        mbox_lines.append(f"Content-Type: text/plain; charset=UTF-8\n")
        mbox_lines.append(f"\n")
        for body_line in email['body'].split('\n'):
            if body_line.startswith('From '):
                body_line = '>' + body_line
            mbox_lines.append(body_line + '\n')
        mbox_lines.append('\n')

    mbox_content = ''.join(mbox_lines)

    with open(finance_path, 'w', encoding='utf-8') as f:
        f.write(mbox_content)
    print(f'Finance mbox written: {finance_path} ({len(FINANCE_EMAILS)} emails)')

    # Remove any stale .msf index so Thunderbird re-indexes
    msf_path = finance_path + '.msf'
    if os.path.exists(msf_path):
        os.remove(msf_path)
        print(f'Removed stale index: {msf_path}')

    # Ensure an Inbox exists in Local Folders (required for Thunderbird to show Local Folders)
    inbox_path = os.path.join(mail_dir, 'Inbox')
    if not os.path.exists(inbox_path):
        with open(inbox_path, 'w', encoding='utf-8') as f:
            f.write('')  # empty inbox
        print(f'Created empty Inbox: {inbox_path}')

    print(f'Finance folder ready. First email subject: {FIRST_EMAIL_SUBJECT}')


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def main():
    create_initial_ods()
    setup_thunderbird_finance_folder()

    # Launch Thunderbird first (agent needs to find Finance folder)
    launch_gui('thunderbird', delay_sec=3.0)

    # Launch LibreOffice Calc with finance_tracker.ods
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Thunderbird and LibreOffice Calc with DISPLAY=:0')


main()
