"""
Reward Script: Track payment confirmation emails
Task ID: osworld_multi_apps_email_data_002
Domain: libreoffice_calc (ODS file — internally XLSX format)

Task: Open Thunderbird Finance folder, find the first email, copy its subject
      line into cell A2 of /home/user/finance_tracker.ods.

Scoring Rubric:
  Component 1: A2 is non-empty (agent wrote something into A2)         — 0.3 pts
  Component 2: A2 contains the exact subject of the first email        — 0.7 pts
               ("Payment Confirmation - Invoice #2024-0847")
  Total: 1.0

Note: finance_tracker.ods is internally an XLSX file.
      We copy it to a .xlsx extension before loading with openpyxl.
"""

import os
import shutil
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_data_002'
ODS_PATH = os.path.join(WORKDIR, 'finance_tracker.ods')

# Ground truth: subject of first email in Thunderbird Finance folder
EXPECTED_SUBJECT = 'Payment Confirmation - Invoice #2024-0847'


def verify_task(ods_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # The file is internally XLSX format despite .ods extension.
    # Copy to /tmp with .xlsx extension so openpyxl can load it.
    tmp_path = '/tmp/finance_tracker_reward_check.xlsx'
    try:
        shutil.copy(ods_path, tmp_path)
    except Exception as e:
        print("CRITICAL: Cannot copy file %s -> %s: %s" % (ods_path, tmp_path, e))
        print("REWARD: 0.0")
        return 0.0

    try:
        wb = openpyxl.load_workbook(tmp_path)
    except Exception as e:
        print("CRITICAL: Cannot load file %s: %s" % (tmp_path, e))
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition gate: verify headers are intact (A1=Subject, B1=Sender, C1=Date)
    try:
        a1 = ws['A1'].value
        b1 = ws['B1'].value
        c1 = ws['C1'].value
        if a1 != 'Subject' or b1 != 'Sender' or c1 != 'Date':
            print("CRITICAL: Headers not intact — A1=%r, B1=%r, C1=%r" % (a1, b1, c1))
            print("REWARD: 0.0")
            return 0.0
        print("GATE PASS: Headers intact (A1=%r, B1=%r, C1=%r)" % (a1, b1, c1))
    except Exception as e:
        print("CRITICAL: Cannot read headers: %s" % e)
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A2 is non-empty — agent wrote something (0.3 points)
    # This FAILS on initial (A2=None) and PASSES on golden (A2 has content)
    try:
        a2_value = ws['A2'].value
        if a2_value is not None and str(a2_value).strip() != '':
            print("PASS: Component 1 — A2 is non-empty (value: %r) (0.3 pts)" % a2_value)
            total_score += 0.3
        else:
            print("FAIL: Component 1 — A2 is empty (expected an email subject in A2)")
    except Exception as e:
        print("ERROR: Component 1 — %s" % e)

    # Component 2: A2 contains the exact subject of the first email (0.7 points)
    # The first email in the Finance folder has subject: "Payment Confirmation - Invoice #2024-0847"
    # This FAILS on initial (A2=None) and PASSES on golden (A2=expected subject)
    try:
        a2_value = ws['A2'].value
        if a2_value is not None:
            a2_str = str(a2_value).strip()
            if a2_str == EXPECTED_SUBJECT:
                print("PASS: Component 2 — A2 matches expected subject exactly: %r (0.7 pts)" % a2_str)
                total_score += 0.7
            else:
                print("FAIL: Component 2 — A2 value %r does not match expected %r" % (a2_str, EXPECTED_SUBJECT))
        else:
            print("FAIL: Component 2 — A2 is None, expected: %r" % EXPECTED_SUBJECT)
    except Exception as e:
        print("ERROR: Component 2 — %s" % e)

    final_score = min(total_score, 1.0)
    print("\nScore: %s/1.0" % total_score)
    print("REWARD: %s" % final_score)
    return final_score


if not os.path.exists(ODS_PATH):
    print("File not found: %s" % ODS_PATH)
    print("REWARD: 0.0")
else:
    verify_task(ODS_PATH)
