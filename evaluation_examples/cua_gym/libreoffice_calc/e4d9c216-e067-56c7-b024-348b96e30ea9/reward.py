"""
Reward Script: Currency Conversion with VLOOKUP, Review Flags, and Conditional Formatting
Task ID: calc_fin_currency_conversion_050
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: VLOOKUP formulas in E2:E50 (USD conversion)          — 0.30 pts
  Component 2: IF review flag formulas in F2:F50 + F1 header        — 0.20 pts
  Component 3: Total row (A51 label + E51 SUM formula, bold)        — 0.20 pts
  Component 4: Number formatting (E column USD, D column decimal)   — 0.15 pts
  Component 5: Conditional formatting on E2:E50 (>10000 yellow)     — 0.15 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_currency_conversion_050'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist
    if 'ForeignInvoices' not in wb.sheetnames:
        print("CRITICAL: 'ForeignInvoices' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ForeignInvoices']

    # -------------------------------------------------------------------------
    # Component 1: VLOOKUP formulas in E2:E50 (0.30 points)
    # The E column must contain =D{row}*VLOOKUP(C{row},ExchangeRates.$A$2:$B$6,2,0)
    # In the initial file, E2:E50 are all empty (None).
    # -------------------------------------------------------------------------
    try:
        vlookup_count = 0
        vlookup_correct_count = 0
        for row in range(2, 51):
            val = ws.cell(row=row, column=5).value  # Column E
            if val is not None:
                vlookup_count += 1
                # Check that formula references VLOOKUP and ExchangeRates
                val_str = str(val).upper().replace(" ", "")
                if 'VLOOKUP' in val_str and 'EXCHANGERATES' in val_str:
                    vlookup_correct_count += 1

        if vlookup_correct_count == 49:
            print(f"PASS: Component 1 — All 49 VLOOKUP formulas in E2:E50 present (0.30 pts)")
            total_score += 0.30
        elif vlookup_correct_count >= 40:
            print(f"PARTIAL: Component 1 — {vlookup_correct_count}/49 VLOOKUP formulas in E column (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {vlookup_correct_count}/49 VLOOKUP formulas found in E2:E50 (found {vlookup_count} non-empty cells)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Review Flag column (F1 header + F2:F50 IF formulas) (0.20 points)
    # F1 should be 'Review Flag', F2:F50 should contain =IF(E{row}>10000,"REVIEW","")
    # In the initial file, F column is entirely empty.
    # -------------------------------------------------------------------------
    try:
        f1_val = ws.cell(row=1, column=6).value  # F1
        f1_ok = f1_val is not None and str(f1_val).strip().lower() in ['review flag', 'reviewflag']

        if_count = 0
        for row in range(2, 51):
            val = ws.cell(row=row, column=6).value  # Column F
            if val is not None:
                val_str = str(val).upper().replace(" ", "")
                if 'IF' in val_str and 'E' in val_str and '10000' in val_str:
                    if_count += 1

        if f1_ok and if_count == 49:
            print(f"PASS: Component 2 — F1='Review Flag' header and all 49 IF formulas in F2:F50 (0.20 pts)")
            total_score += 0.20
        elif if_count >= 40:
            partial = 0.10
            if f1_ok:
                partial = 0.12
            print(f"PARTIAL: Component 2 — F1={'OK' if f1_ok else 'missing'}, {if_count}/49 IF formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — F1={repr(f1_val)}, only {if_count}/49 IF formulas in F column")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Total row A51 + E51 (SUM formula, bold) (0.20 points)
    # A51: 'Total USD Exposure', E51: =SUM(E2:E50), both bold
    # In the initial file, row 51 is entirely empty.
    # -------------------------------------------------------------------------
    try:
        a51_val = ws.cell(row=51, column=1).value  # A51
        e51_val = ws.cell(row=51, column=5).value  # E51
        a51_bold = ws.cell(row=51, column=1).font.bold
        e51_bold = ws.cell(row=51, column=5).font.bold

        a51_ok = a51_val is not None and 'total' in str(a51_val).lower()
        e51_ok = e51_val is not None and 'SUM' in str(e51_val).upper() and 'E2' in str(e51_val).upper()

        if a51_ok and e51_ok and a51_bold and e51_bold:
            print(f"PASS: Component 3 — A51='{a51_val}', E51='{e51_val}', both bold (0.20 pts)")
            total_score += 0.20
        elif a51_ok and e51_ok:
            print(f"PARTIAL: Component 3 — A51 and E51 values correct but bold check: A51_bold={a51_bold}, E51_bold={e51_bold} (0.10 pts)")
            total_score += 0.10
        elif e51_ok:
            print(f"PARTIAL: Component 3 — E51 SUM formula present, A51 label missing/wrong (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — A51={repr(a51_val)}, E51={repr(e51_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Number formatting (E column USD, D column 2 decimal) (0.15 points)
    # E2:E50 should be formatted as USD: '$#,##0.00'
    # D2:D50 should be formatted with 2 decimal places: '#,##0.00'
    # In the initial file, D is 'General' format and E is empty.
    # -------------------------------------------------------------------------
    try:
        e_fmt_count = 0
        d_fmt_count = 0
        for row in range(2, 51):
            e_cell = ws.cell(row=row, column=5)
            d_cell = ws.cell(row=row, column=4)
            # E column: must include $ or USD format
            e_fmt = e_cell.number_format or ''
            if '$' in e_fmt or 'USD' in e_fmt.upper():
                e_fmt_count += 1
            # D column: must have decimal formatting (not 'General')
            d_fmt = d_cell.number_format or ''
            if d_fmt != 'General' and '0.00' in d_fmt:
                d_fmt_count += 1

        e_fmt_ok = e_fmt_count >= 45  # most E cells have USD format
        d_fmt_ok = d_fmt_count >= 45  # most D cells have decimal format

        if e_fmt_ok and d_fmt_ok:
            print(f"PASS: Component 4 — E column USD format ({e_fmt_count}/49 cells), D column decimal format ({d_fmt_count}/49 cells) (0.15 pts)")
            total_score += 0.15
        elif e_fmt_ok:
            print(f"PARTIAL: Component 4 — E column USD format OK ({e_fmt_count}/49), D column format incomplete ({d_fmt_count}/49) (0.08 pts)")
            total_score += 0.08
        elif d_fmt_ok:
            print(f"PARTIAL: Component 4 — D column decimal format OK ({d_fmt_count}/49), E column USD format missing ({e_fmt_count}/49) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — E USD format: {e_fmt_count}/49, D decimal format: {d_fmt_count}/49")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting on E2:E50 (>10000 → yellow) (0.15 points)
    # A CellIs rule greaterThan 10000 with a yellow fill must be applied to E2:E50.
    # In the initial file, no conditional formatting exists on E column.
    # -------------------------------------------------------------------------
    try:
        cf_found = False
        cf_yellow = False
        cf_rules = ws.conditional_formatting
        for cf in cf_rules:
            cf_str = str(cf)
            # Check if it covers the E column range
            if 'E2' in cf_str or 'E2:E50' in cf_str:
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', '')
                    op = getattr(rule, 'operator', '')
                    formula = getattr(rule, 'formula', [])
                    if rule_type == 'cellIs' and op == 'greaterThan' and formula and '10000' in str(formula[0]):
                        cf_found = True
                        # Check for yellow fill
                        dxf = getattr(rule, 'dxf', None)
                        if dxf and dxf.fill:
                            try:
                                fg_rgb = dxf.fill.fgColor.rgb
                                # Yellow variants: FFFFFF00, FFFF00 (6-char), pure yellow
                                if fg_rgb and ('FF00' in fg_rgb.upper() or fg_rgb.upper() in ['FFFFFF00', 'FFFF0000']):
                                    cf_yellow = True
                                # More permissive yellow check: ends with FF00 (RR=FF, GG=FF, BB=00)
                                if fg_rgb and len(fg_rgb) >= 6:
                                    # ARGB: last 6 chars are RRGGBB
                                    rrggbb = fg_rgb[-6:].upper()
                                    if rrggbb == 'FFFF00':
                                        cf_yellow = True
                            except Exception:
                                pass

        if cf_found and cf_yellow:
            print(f"PASS: Component 5 — Conditional formatting on E2:E50 >10000 with yellow fill (0.15 pts)")
            total_score += 0.15
        elif cf_found:
            print(f"PARTIAL: Component 5 — Conditional formatting rule found but yellow fill not confirmed (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 5 — No conditional formatting rule found on E2:E50 with >10000 condition")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
