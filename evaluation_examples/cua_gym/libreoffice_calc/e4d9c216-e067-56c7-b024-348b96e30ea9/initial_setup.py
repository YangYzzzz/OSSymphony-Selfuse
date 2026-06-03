"""
Initial Setup: Currency conversion invoices spreadsheet
Task ID: calc_fin_currency_conversion_050
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_currency_conversion_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: ForeignInvoices ---
    ws1 = wb.active
    ws1.title = 'ForeignInvoices'

    # Headers (row 1) — NOT bold in initial state
    headers = ['Invoice#', 'Vendor', 'Currency', 'Foreign Amount', 'USD Amount']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)
    # Note: F1 'Review Flag' will be added in golden patch

    # Invoice data - realistic vendor names, currencies, amounts
    # Currencies: EUR, GBP, JPY, CAD, AUD
    invoice_data = [
        ('INV-2025-001', 'Müller GmbH',          'EUR',  8750.00),
        ('INV-2025-002', 'TechVision Ltd',        'GBP', 12400.00),
        ('INV-2025-003', 'Sakura Electronics',    'JPY', 1850000.00),
        ('INV-2025-004', 'Northern Supplies Co',  'CAD', 15200.00),
        ('INV-2025-005', 'Sydney Partners Pty',   'AUD', 22600.00),
        ('INV-2025-006', 'Bergmann & Söhne',      'EUR',  3420.50),
        ('INV-2025-007', 'Pinnacle Solutions UK',  'GBP',  9870.00),
        ('INV-2025-008', 'Nippon Trading Corp',   'JPY', 2340000.00),
        ('INV-2025-009', 'Maple Leaf Industries', 'CAD',  6750.00),
        ('INV-2025-010', 'AusTech Innovations',   'AUD', 11300.00),
        ('INV-2025-011', 'Schneider Logistics',   'EUR', 17500.00),
        ('INV-2025-012', 'Harrington & Wells',    'GBP',  4580.00),
        ('INV-2025-013', 'Fuji Components Ltd',   'JPY',  890000.00),
        ('INV-2025-014', 'Ontario Packaging Inc', 'CAD', 23400.00),
        ('INV-2025-015', 'Brisbane Retail Group', 'AUD',  7890.00),
        ('INV-2025-016', 'Frankfurter Systeme',   'EUR',  5640.00),
        ('INV-2025-017', 'Whitmore Consulting',   'GBP', 18900.00),
        ('INV-2025-018', 'Tokyo Digital Works',   'JPY', 4200000.00),
        ('INV-2025-019', 'Vancouver Marine Ltd',  'CAD',  9100.00),
        ('INV-2025-020', 'Melbourne Holdings',    'AUD', 31000.00),
        ('INV-2025-021', 'Braun Elektronik AG',   'EUR', 11200.00),
        ('INV-2025-022', 'Stratford & Holt',      'GBP',  6730.00),
        ('INV-2025-023', 'Osaka Precision Mfg',   'JPY', 1120000.00),
        ('INV-2025-024', 'Calgary Energy Corp',   'CAD', 18700.00),
        ('INV-2025-025', 'Perth Mining Services', 'AUD', 14500.00),
        ('INV-2025-026', 'Vogel Industrietechnik','EUR',  2980.00),
        ('INV-2025-027', 'Lancaster Textiles',    'GBP', 11560.00),
        ('INV-2025-028', 'Kyoto Craft Studios',   'JPY',  670000.00),
        ('INV-2025-029', 'Halifax Seafood Ltd',   'CAD',  4320.00),
        ('INV-2025-030', 'Adelaide Wineries',     'AUD',  8950.00),
        ('INV-2025-031', 'Keller & Partner',      'EUR', 24500.00),
        ('INV-2025-032', 'Bromley Engineering',   'GBP',  3190.00),
        ('INV-2025-033', 'Hiroshima Metal Works', 'JPY', 3780000.00),
        ('INV-2025-034', 'Winnipeg Grain Corp',   'CAD', 12600.00),
        ('INV-2025-035', 'Darwin Transport Co',   'AUD', 19200.00),
        ('INV-2025-036', 'Heidelberg Press AG',   'EUR',  7320.00),
        ('INV-2025-037', 'Cambridge Analytics',   'GBP', 21400.00),
        ('INV-2025-038', 'Sapporo Beverage Ltd',  'JPY', 5600000.00),
        ('INV-2025-039', 'Quebec Pharma Inc',     'CAD',  3870.00),
        ('INV-2025-040', 'Hobart Fisheries Ltd',  'AUD',  6100.00),
        ('INV-2025-041', 'Darmstadt Chemicals',   'EUR', 33100.00),
        ('INV-2025-042', 'Durham Construction',   'GBP',  8450.00),
        ('INV-2025-043', 'Yokohama Shipbuilding', 'JPY', 9200000.00),
        ('INV-2025-044', 'Saskatoon Agri Corp',   'CAD', 27800.00),
        ('INV-2025-045', 'Gold Coast Resorts',    'AUD', 41500.00),
        ('INV-2025-046', 'Augsburg Autotech',     'EUR', 15600.00),
        ('INV-2025-047', 'Bristol Aerospace',     'GBP', 29700.00),
        ('INV-2025-048', 'Nagoya Auto Parts',     'JPY', 7100000.00),
        ('INV-2025-049', 'Thunder Bay Lumber',    'CAD',  5490.00),
    ]

    for r, (inv, vendor, currency, amount) in enumerate(invoice_data, 2):
        ws1.cell(row=r, column=1, value=inv)
        ws1.cell(row=r, column=2, value=vendor)
        ws1.cell(row=r, column=3, value=currency)
        ws1.cell(row=r, column=4, value=amount)
        # Column E (USD Amount) left empty — to be filled by task
        # Column F (Review Flag) not added yet

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 26
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 16

    # --- Sheet 2: ExchangeRates ---
    ws2 = wb.create_sheet('ExchangeRates')

    ws2['A1'] = 'Currency'
    ws2['B1'] = 'Rate to USD'

    exchange_data = [
        ('EUR', 1.08),
        ('GBP', 1.27),
        ('JPY', 0.0067),
        ('CAD', 0.74),
        ('AUD', 0.65),
    ]

    for r, (currency, rate) in enumerate(exchange_data, 2):
        ws2.cell(row=r, column=1, value=currency)
        ws2.cell(row=r, column=2, value=rate)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
