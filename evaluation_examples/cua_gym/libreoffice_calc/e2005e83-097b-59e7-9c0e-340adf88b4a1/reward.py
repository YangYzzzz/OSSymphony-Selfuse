"""
Reward Script: Filter product ratings table to show only below-average rated products
Task ID: calc_dop_filter_belowavg_054
Domain: libreoffice_calc
Scoring:
  Component 1: AutoFilter on column D has a belowAverage dynamic filter (0.5 pts)
  Component 2: Rows with rating >= average are hidden (0.3 pts)
  Component 3: Below-avg rows visible AND above-avg rows hidden (compound check) (0.2 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_filter_belowavg_054'
SHEET_NAME = 'ProductRatings'
DATA_ROWS = range(2, 62)  # rows 2-61, 60 data rows
RATING_COL = 4            # Column D


def _has_below_avg_dynamic_filter(ws):
    """
    Check whether ws.auto_filter has a DynamicFilter with type='belowAverage'
    on colId=3 (column D, 0-indexed within the filter range).
    Returns (found: bool, detail: str).
    """
    filter_columns = ws.auto_filter.filterColumn
    for fc in filter_columns:
        if fc.colId == 3:
            df = fc.dynamicFilter
            if df is not None and df.type == 'belowAverage':
                return True, f"val={df.val:.4f}"
    found_cols = [(fc.colId, str(fc.dynamicFilter)) for fc in filter_columns]
    return False, f"filterColumns={found_cols}"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Filter the 'ProductRatings' sheet so only rows with Rating (column D)
    below the sheet average are visible.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — fail fast if file is corrupt/missing
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Compute the average rating from all data rows (using the file's own data)
    ratings = []
    for row in DATA_ROWS:
        val = ws.cell(row=row, column=RATING_COL).value
        try:
            ratings.append(float(val))
        except (TypeError, ValueError):
            pass

    if not ratings:
        print("CRITICAL: No rating data found in column D.")
        print("REWARD: 0.0")
        return 0.0

    avg_rating = sum(ratings) / len(ratings)
    print(f"INFO: Computed average rating = {avg_rating:.4f} from {len(ratings)} rows")

    # ----------------------------------------------------------------
    # Component 1: AutoFilter on column D has a belowAverage dynamic filter (0.5 pts)
    # The filter was applied on column D (colId=3, 0-indexed) using Standard Filter
    # or the 'Below Average' option. This introduces a DynamicFilter with type='belowAverage'.
    # The initial file has no filterColumn entries (filterColumn list is empty).
    # The golden file has one filterColumn entry with colId=3 and dynamicFilter.type='belowAverage'.
    # ----------------------------------------------------------------
    try:
        found, detail = _has_below_avg_dynamic_filter(ws)
        if found:
            print(f"PASS: Component 1 — AutoFilter on column D has belowAverage "
                  f"dynamic filter ({detail}) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Expected AutoFilter column D (colId=3) "
                  f"with belowAverage dynamic filter. Found: {detail}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: Rows with rating >= average are hidden (0.3 pts)
    # All products whose rating is at or above the computed average should be hidden.
    # In the initial file, no rows are hidden (0 hidden). In the golden file, all 31
    # above-avg rows are hidden via the auto-filter mechanism.
    # ----------------------------------------------------------------
    try:
        above_avg_rows = [r for r in DATA_ROWS
                          if ws.cell(row=r, column=RATING_COL).value is not None
                          and float(ws.cell(row=r, column=RATING_COL).value) >= avg_rating]

        if not above_avg_rows:
            print("FAIL: Component 2 — No rows found with rating >= average")
        else:
            hidden_correctly = sum(
                1 for r in above_avg_rows
                if ws.row_dimensions.get(r) and ws.row_dimensions[r].hidden
            )
            fraction_hidden = hidden_correctly / len(above_avg_rows)
            if fraction_hidden >= 1.0:
                print(f"PASS: Component 2 — All {len(above_avg_rows)} rows with "
                      f"rating >= avg are hidden (0.3 pts)")
                total_score += 0.3
            elif fraction_hidden >= 0.8:
                partial = round(0.3 * fraction_hidden, 2)
                total_score += partial
                print(f"PARTIAL: Component 2 — {hidden_correctly}/{len(above_avg_rows)} "
                      f"above-avg rows hidden ({partial:.2f} pts)")
            else:
                print(f"FAIL: Component 2 — Only {hidden_correctly}/{len(above_avg_rows)} "
                      f"above-avg rows are hidden (expected all {len(above_avg_rows)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Below-avg rows are visible AND above-avg rows are hidden (0.2 pts)
    # This is a compound check — both conditions must hold simultaneously.
    # In the initial file, no rows are hidden so above-avg rows are NOT hidden,
    # making this component FAIL on the initial file (correct — it only passes post-filter).
    # In the golden file, both conditions hold after the filter has been applied.
    # ----------------------------------------------------------------
    try:
        below_avg_rows = [r for r in DATA_ROWS
                          if ws.cell(row=r, column=RATING_COL).value is not None
                          and float(ws.cell(row=r, column=RATING_COL).value) < avg_rating]
        above_avg_rows_c3 = [r for r in DATA_ROWS
                              if ws.cell(row=r, column=RATING_COL).value is not None
                              and float(ws.cell(row=r, column=RATING_COL).value) >= avg_rating]

        if not below_avg_rows or not above_avg_rows_c3:
            print("FAIL: Component 3 — Insufficient data for compound visibility check")
        else:
            visible_below = sum(
                1 for r in below_avg_rows
                if not (ws.row_dimensions.get(r) and ws.row_dimensions[r].hidden)
            )
            hidden_above = sum(
                1 for r in above_avg_rows_c3
                if ws.row_dimensions.get(r) and ws.row_dimensions[r].hidden
            )
            # Both must hold: all below-avg rows visible AND all above-avg rows hidden
            compound_ok = (visible_below == len(below_avg_rows)
                           and hidden_above == len(above_avg_rows_c3))
            if compound_ok:
                print(f"PASS: Component 3 — All {len(below_avg_rows)} below-avg rows visible "
                      f"AND all {len(above_avg_rows_c3)} above-avg rows hidden (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Compound check failed: "
                      f"below-avg visible={visible_below}/{len(below_avg_rows)}, "
                      f"above-avg hidden={hidden_above}/{len(above_avg_rows_c3)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
