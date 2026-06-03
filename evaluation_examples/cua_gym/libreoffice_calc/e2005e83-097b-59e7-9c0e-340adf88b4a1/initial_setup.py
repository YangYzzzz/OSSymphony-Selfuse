"""
Initial Setup: Filter product ratings table to show below-average products
Task ID: calc_dop_filter_belowavg_054
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.worksheet.filters import AutoFilter

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_filter_belowavg_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProductRatings ---
    ws = wb.active
    ws.title = 'ProductRatings'

    # Headers
    headers = ['Product ID', 'Product Name', 'Category', 'Rating', 'Review Count', 'Sales Rank']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 60 rows of realistic product data
    # Ratings span 2.1 to 4.9 with average ~3.7
    # Ratings below 3.7: products are "underperforming"
    data = [
        # (Product ID, Product Name, Category, Rating, Review Count, Sales Rank)
        ('P001', 'Budget Bluetooth Speaker', 'Electronics', 2.8, 145, 1240),
        ('P002', 'Premium Headphones', 'Electronics', 4.6, 892, 34),
        ('P003', 'Wireless Mouse', 'Electronics', 3.9, 534, 210),
        ('P004', 'USB-C Hub 7-Port', 'Electronics', 3.2, 278, 560),
        ('P005', 'Ergonomic Keyboard', 'Electronics', 4.4, 1203, 67),
        ('P006', 'Portable Charger 20000mAh', 'Electronics', 3.5, 421, 320),
        ('P007', 'Smart LED Desk Lamp', 'Home & Office', 4.1, 367, 185),
        ('P008', 'Stainless Steel Water Bottle', 'Sports', 4.7, 2145, 12),
        ('P009', 'Foam Yoga Mat', 'Sports', 3.3, 189, 740),
        ('P010', 'Resistance Band Set', 'Sports', 4.2, 876, 95),
        ('P011', 'Non-Stick Frying Pan', 'Kitchen', 3.6, 654, 275),
        ('P012', 'French Press Coffee Maker', 'Kitchen', 4.5, 1087, 52),
        ('P013', 'Adjustable Dumbbell Pair', 'Sports', 2.9, 98, 1580),
        ('P014', 'Laptop Stand Aluminum', 'Electronics', 4.3, 723, 118),
        ('P015', 'Mechanical Pencil Set', 'Office Supplies', 3.1, 234, 820),
        ('P016', 'Noise Cancelling Earbuds', 'Electronics', 4.8, 3421, 8),
        ('P017', 'Canvas Backpack 30L', 'Travel', 3.4, 445, 390),
        ('P018', 'Stainless Mixing Bowls Set', 'Kitchen', 4.0, 512, 195),
        ('P019', 'Digital Kitchen Scale', 'Kitchen', 3.7, 387, 302),
        ('P020', 'Bamboo Cutting Board', 'Kitchen', 2.6, 112, 1450),
        ('P021', 'Indoor Plant Pot Set', 'Home & Garden', 4.2, 298, 148),
        ('P022', 'Desk Organizer Tray', 'Office Supplies', 3.0, 176, 870),
        ('P023', 'HDMI Cable 6ft 2-Pack', 'Electronics', 4.4, 1654, 43),
        ('P024', 'Anti-Fatigue Standing Mat', 'Home & Office', 3.8, 421, 255),
        ('P025', 'Reusable Grocery Bags Set', 'Home & Garden', 4.6, 987, 27),
        ('P026', 'Ceramic Coffee Mug 16oz', 'Kitchen', 2.4, 67, 1780),
        ('P027', 'Wooden Chess Set', 'Toys & Games', 3.9, 234, 228),
        ('P028', 'Memory Foam Pillow', 'Bedding', 4.1, 876, 102),
        ('P029', 'Travel Neck Pillow', 'Travel', 3.2, 345, 495),
        ('P030', 'Hanging Shower Organizer', 'Home & Garden', 2.7, 134, 1320),
        ('P031', 'Electric Toothbrush', 'Personal Care', 4.5, 2187, 38),
        ('P032', 'Sunscreen SPF 50 2-Pack', 'Personal Care', 3.6, 423, 310),
        ('P033', 'Insulated Lunch Bag', 'Travel', 4.3, 765, 86),
        ('P034', 'Phone Stand Adjustable', 'Electronics', 3.1, 312, 640),
        ('P035', 'Smart Plug Wi-Fi 4-Pack', 'Smart Home', 4.7, 4532, 15),
        ('P036', 'Door Stopper Heavy Duty', 'Home & Garden', 2.5, 89, 1670),
        ('P037', 'Yoga Blocks 2-Pack', 'Sports', 4.0, 543, 167),
        ('P038', 'Bamboo Fiber Towel Set', 'Bedding', 3.4, 267, 455),
        ('P039', 'Essential Oil Diffuser', 'Home & Garden', 4.2, 1098, 73),
        ('P040', 'Beeswax Wrap Set', 'Kitchen', 3.3, 198, 698),
        ('P041', 'Foldable Tote Bag', 'Travel', 4.4, 876, 55),
        ('P042', 'Silicone Baking Mats 2-Pack', 'Kitchen', 4.6, 1234, 31),
        ('P043', 'Collapsible Water Bottle', 'Sports', 3.5, 356, 378),
        ('P044', 'Bluetooth Sleep Mask', 'Personal Care', 2.3, 78, 1890),
        ('P045', 'Laptop Sleeve 15.6 inch', 'Electronics', 3.8, 489, 248),
        ('P046', 'Magnetic Whiteboard Markers', 'Office Supplies', 4.1, 678, 122),
        ('P047', 'Spice Rack Organizer', 'Kitchen', 3.6, 312, 318),
        ('P048', 'Microfiber Cleaning Cloths', 'Home & Garden', 4.3, 2341, 62),
        ('P049', 'Cable Management Box', 'Home & Office', 3.0, 245, 835),
        ('P050', 'Digital Alarm Clock', 'Home & Office', 2.9, 134, 1220),
        ('P051', 'Compression Packing Cubes', 'Travel', 4.5, 1567, 41),
        ('P052', 'Wood Grain Desk Pad', 'Office Supplies', 3.7, 423, 283),
        ('P053', 'Mesh Laundry Bags 6-Pack', 'Home & Garden', 4.0, 789, 158),
        ('P054', 'Portable Mini Projector', 'Electronics', 3.2, 167, 620),
        ('P055', 'Herb Keeper Container', 'Kitchen', 2.1, 56, 2100),
        ('P056', 'Wooden Wall Clock', 'Home & Garden', 4.8, 1876, 9),
        ('P057', 'Resistance Loop Bands', 'Sports', 3.9, 654, 198),
        ('P058', 'Leather Passport Holder', 'Travel', 4.9, 2134, 5),
        ('P059', 'Gel Wrist Rest Pad', 'Office Supplies', 3.4, 387, 415),
        ('P060', 'Aromatherapy Shower Tablets', 'Personal Care', 2.8, 112, 1380),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Enable AutoFilter on row 1 (no filters applied - all rows visible)
    ws.auto_filter.ref = 'A1:F61'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Print the average rating for reference
    ratings = [row[3] for row in data]
    avg = sum(ratings) / len(ratings)
    below_avg = [r for r in data if r[3] < avg]
    print(f'Total rows: {len(data)}')
    print(f'Average rating: {avg:.4f}')
    print(f'Rows below average: {len(below_avg)}')
    print(f'Rows at or above average: {len(data) - len(below_avg)}')


create_initial()
