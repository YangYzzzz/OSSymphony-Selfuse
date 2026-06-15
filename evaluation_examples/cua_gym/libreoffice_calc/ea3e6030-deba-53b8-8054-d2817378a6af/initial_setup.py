"""
Initial Setup: Create Code_Registry spreadsheet with 34 entries, no validation
Task ID: calc_gcv_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Code_Registry"

    # Headers
    headers = ["Sequence", "Description", "Code"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic descriptions for 34 entries across various registry categories
    descriptions = [
        "Primary warehouse location identifier",
        "Customer loyalty tier classification",
        "Shipping carrier service level",
        "Product category designation",
        "Regional distribution center tag",
        "Quality assurance batch marker",
        "Employee department assignment",
        "Vendor compliance certification",
        "Equipment maintenance priority",
        "Invoice processing workflow stage",
        "Hazardous material handling class",
        "Return authorization status flag",
        "Inventory reorder threshold level",
        "Customs clearance document type",
        "Fleet vehicle assignment group",
        "Safety inspection certification grade",
        "Budget allocation department code",
        "Telecommunications routing prefix",
        "Patient care priority indicator",
        "Environmental compliance category",
        "Building access security clearance",
        "Procurement approval authority tier",
        "Data classification sensitivity label",
        "Training completion certification mark",
        "Emergency response escalation level",
        "Payroll deduction category identifier",
        "Asset depreciation schedule class",
        "Network segment isolation zone tag",
        "Pharmaceutical storage condition type",
        "Project milestone tracking phase",
        "Legal document retention category",
        "Manufacturing line assignment sector",
        "Credit risk assessment rating band",
        "Supply chain disruption severity index",
    ]

    for i in range(34):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)               # Sequence
        ws.cell(row=row, column=2, value=descriptions[i])      # Description
        # Column C (Code) intentionally left EMPTY — task is to add validation

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 15

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
