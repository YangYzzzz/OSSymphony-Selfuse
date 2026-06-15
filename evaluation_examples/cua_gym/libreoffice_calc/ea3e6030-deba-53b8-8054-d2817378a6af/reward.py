"""
Reward Script: Data validation on C2:C35 restricting to 2-10 uppercase letters
Task ID: calc_gcv_085
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Custom data validation exists on C2:C35
  Component 2 (0.35): Formula matches =AND(EXACT(C2,UPPER(C2)),LEN(C2)>=2,LEN(C2)<=10)
  Component 3 (0.35): Error alert configured (stop, title, message, showErrorMessage)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_085'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: sheet is Code_Registry
    if ws.title != 'Code_Registry':
        # Try to find it
        if 'Code_Registry' in wb.sheetnames:
            ws = wb['Code_Registry']
        else:
            print(f"FAIL: Sheet 'Code_Registry' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Get data validations
    dvs = ws.data_validations.dataValidation if ws.data_validations else []

    # Find the relevant data validation targeting C2:C35 (or overlapping range in column C)
    target_dv = None
    for dv in dvs:
        sqref_str = str(dv.sqref).upper()
        # Check if the validation covers column C in the expected range
        if 'C2' in sqref_str and 'C35' in sqref_str:
            target_dv = dv
            break
    # Fallback: find any custom validation on column C range
    if target_dv is None:
        for dv in dvs:
            sqref_str = str(dv.sqref).upper()
            if 'C' in sqref_str and dv.type == 'custom':
                target_dv = dv
                break

    # Component 1: Custom data validation exists on C2:C35 (0.30 points)
    try:
        if target_dv is not None and target_dv.type == 'custom':
            sqref_str = str(target_dv.sqref).upper().replace(' ', '')
            if 'C2:C35' in sqref_str or sqref_str == 'C2:C35':
                print(f"PASS: Component 1 — Custom validation on C2:C35 (0.30 pts)")
                total_score += 0.30
            else:
                print(f"PARTIAL: Component 1 — Custom validation found but range is {target_dv.sqref}, expected C2:C35 (0.15 pts)")
                total_score += 0.15
        else:
            if target_dv is not None:
                print(f"FAIL: Component 1 — Validation found but type is '{target_dv.type}', expected 'custom'")
            else:
                print(f"FAIL: Component 1 — No data validation found on C2:C35")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula matches =AND(EXACT(C2,UPPER(C2)),LEN(C2)>=2,LEN(C2)<=10) (0.35 points)
    try:
        if target_dv is not None and target_dv.formula1:
            formula = str(target_dv.formula1).strip()
            # Normalize for comparison: uppercase, remove spaces
            norm_formula = formula.upper().replace(' ', '').replace('=', '', 1) if formula.startswith('=') else formula.upper().replace(' ', '')
            norm_formula_with_eq = '=' + norm_formula

            expected_core = 'AND(EXACT(C2,UPPER(C2)),LEN(C2)>=2,LEN(C2)<=10)'
            expected_norm = expected_core.upper().replace(' ', '')

            # Check full formula
            actual_check = norm_formula if not formula.startswith('=') else norm_formula
            actual_check_clean = formula.upper().replace(' ', '')
            if actual_check_clean.lstrip('=') == expected_norm:
                print(f"PASS: Component 2 — Formula matches: {formula} (0.35 pts)")
                total_score += 0.35
            else:
                # Check if it at least contains EXACT and UPPER and LEN
                has_exact = 'EXACT' in actual_check_clean
                has_upper = 'UPPER' in actual_check_clean
                has_len = 'LEN' in actual_check_clean
                if has_exact and has_upper and has_len:
                    print(f"PARTIAL: Component 2 — Formula has key functions but doesn't match exactly: {formula} (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 2 — Formula: {formula}, expected =AND(EXACT(C2,UPPER(C2)),LEN(C2)>=2,LEN(C2)<=10)")
        else:
            print(f"FAIL: Component 2 — No formula found in data validation")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Error alert configuration (0.35 points)
    # Error style=stop, title='Invalid Code', message='Code must be 2-10 uppercase letters.', showErrorMessage=True
    try:
        if target_dv is not None:
            sub_score = 0.0

            # Error style = stop
            if target_dv.errorStyle and target_dv.errorStyle.lower() == 'stop':
                sub_score += 0.10
                print(f"  PASS: Error style is 'stop'")
            else:
                print(f"  FAIL: Error style is '{target_dv.errorStyle}', expected 'stop'")

            # Error title = 'Invalid Code'
            if target_dv.errorTitle and 'invalid code' in target_dv.errorTitle.lower():
                sub_score += 0.08
                print(f"  PASS: Error title is '{target_dv.errorTitle}'")
            else:
                print(f"  FAIL: Error title is '{target_dv.errorTitle}', expected 'Invalid Code'")

            # Error message contains key info
            if target_dv.error and '2' in target_dv.error and '10' in target_dv.error and 'uppercase' in target_dv.error.lower():
                sub_score += 0.09
                print(f"  PASS: Error message: '{target_dv.error}'")
            else:
                print(f"  FAIL: Error message is '{target_dv.error}', expected something about 2-10 uppercase letters")

            # showErrorMessage = True
            if target_dv.showErrorMessage:
                sub_score += 0.08
                print(f"  PASS: showErrorMessage is True")
            else:
                print(f"  FAIL: showErrorMessage is {target_dv.showErrorMessage}, expected True")

            if sub_score > 0:
                print(f"PASS: Component 3 — Error alert config ({sub_score:.2f} pts)")
                total_score += sub_score
            else:
                print(f"FAIL: Component 3 — No error alert properties matched")
        else:
            print(f"FAIL: Component 3 — No data validation found to check error alert")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
