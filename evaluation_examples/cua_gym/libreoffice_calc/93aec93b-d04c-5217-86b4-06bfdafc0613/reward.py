"""
Reward Script: Build an attendance summary showing days present, absent, and late for each employee.
Task ID: calc_hr_024
Domain: libreoffice_calc
Scoring:
  Component 1 (0.34): Alice's COUNTIF formulas in W2:Y2
  Component 2 (0.33): Bob's COUNTIF formulas in W3:Y3
  Component 3 (0.33): Carol's COUNTIF formulas in W4:Y4
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_024'


def check_countif_formula(ws, cell_coord, expected_row, expected_criteria):
    """
    Check if a cell contains a COUNTIF formula counting the given criteria
    in the attendance range for the given row.
    Returns True if a valid COUNTIF formula is found.
    """
    val = ws[cell_coord].value
    if val is None:
        return False
    if not isinstance(val, str):
        return False
    # Normalize: uppercase, strip spaces
    normalized = val.upper().replace(" ", "")
    # Expected pattern: =COUNTIF(B<row>:V<row>,"<criteria>")
    expected = f'=COUNTIF(B{expected_row}:V{expected_row},"{expected_criteria}")'.upper().replace(" ", "")
    return normalized == expected


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Attendance' sheet exists
    if 'Attendance' not in wb.sheetnames:
        print("CRITICAL: 'Attendance' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Attendance']

    # Component 1: Alice's COUNTIF formulas in W2, X2, Y2 (0.34 points)
    try:
        alice_pass = 0
        # W2: =COUNTIF(B2:V2,"P")
        if check_countif_formula(ws, 'W2', 2, 'P'):
            print("PASS: W2 contains correct COUNTIF for Present (Alice)")
            alice_pass += 1
        else:
            print(f"FAIL: W2 expected COUNTIF for Present, found: {ws['W2'].value}")

        # X2: =COUNTIF(B2:V2,"A")
        if check_countif_formula(ws, 'X2', 2, 'A'):
            print("PASS: X2 contains correct COUNTIF for Absent (Alice)")
            alice_pass += 1
        else:
            print(f"FAIL: X2 expected COUNTIF for Absent, found: {ws['X2'].value}")

        # Y2: =COUNTIF(B2:V2,"L")
        if check_countif_formula(ws, 'Y2', 2, 'L'):
            print("PASS: Y2 contains correct COUNTIF for Late (Alice)")
            alice_pass += 1
        else:
            print(f"FAIL: Y2 expected COUNTIF for Late, found: {ws['Y2'].value}")

        if alice_pass == 3:
            print(f"PASS: Component 1 -- Alice's formulas all correct (0.34 pts)")
            total_score += 0.34
        elif alice_pass > 0:
            partial = round(0.34 * alice_pass / 3, 2)
            print(f"PARTIAL: Component 1 -- Alice has {alice_pass}/3 correct formulas ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 1 -- No correct formulas for Alice")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Bob's COUNTIF formulas in W3, X3, Y3 (0.33 points)
    try:
        bob_pass = 0
        # W3: =COUNTIF(B3:V3,"P")
        if check_countif_formula(ws, 'W3', 3, 'P'):
            print("PASS: W3 contains correct COUNTIF for Present (Bob)")
            bob_pass += 1
        else:
            print(f"FAIL: W3 expected COUNTIF for Present, found: {ws['W3'].value}")

        # X3: =COUNTIF(B3:V3,"A")
        if check_countif_formula(ws, 'X3', 3, 'A'):
            print("PASS: X3 contains correct COUNTIF for Absent (Bob)")
            bob_pass += 1
        else:
            print(f"FAIL: X3 expected COUNTIF for Absent, found: {ws['X3'].value}")

        # Y3: =COUNTIF(B3:V3,"L")
        if check_countif_formula(ws, 'Y3', 3, 'L'):
            print("PASS: Y3 contains correct COUNTIF for Late (Bob)")
            bob_pass += 1
        else:
            print(f"FAIL: Y3 expected COUNTIF for Late, found: {ws['Y3'].value}")

        if bob_pass == 3:
            print(f"PASS: Component 2 -- Bob's formulas all correct (0.33 pts)")
            total_score += 0.33
        elif bob_pass > 0:
            partial = round(0.33 * bob_pass / 3, 2)
            print(f"PARTIAL: Component 2 -- Bob has {bob_pass}/3 correct formulas ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 2 -- No correct formulas for Bob")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Carol's COUNTIF formulas in W4, X4, Y4 (0.33 points)
    try:
        carol_pass = 0
        # W4: =COUNTIF(B4:V4,"P")
        if check_countif_formula(ws, 'W4', 4, 'P'):
            print("PASS: W4 contains correct COUNTIF for Present (Carol)")
            carol_pass += 1
        else:
            print(f"FAIL: W4 expected COUNTIF for Present, found: {ws['W4'].value}")

        # X4: =COUNTIF(B4:V4,"A")
        if check_countif_formula(ws, 'X4', 4, 'A'):
            print("PASS: X4 contains correct COUNTIF for Absent (Carol)")
            carol_pass += 1
        else:
            print(f"FAIL: X4 expected COUNTIF for Absent, found: {ws['X4'].value}")

        # Y4: =COUNTIF(B4:V4,"L")
        if check_countif_formula(ws, 'Y4', 4, 'L'):
            print("PASS: Y4 contains correct COUNTIF for Late (Carol)")
            carol_pass += 1
        else:
            print(f"FAIL: Y4 expected COUNTIF for Late, found: {ws['Y4'].value}")

        if carol_pass == 3:
            print(f"PASS: Component 3 -- Carol's formulas all correct (0.33 pts)")
            total_score += 0.33
        elif carol_pass > 0:
            partial = round(0.33 * carol_pass / 3, 2)
            print(f"PARTIAL: Component 3 -- Carol has {carol_pass}/3 correct formulas ({partial} pts)")
            total_score += partial
        else:
            print("FAIL: Component 3 -- No correct formulas for Carol")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved GUI edits before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
