"""
Initial Setup: Build attendance sheet with daily P/A/L data for 3 employees.
Task ID: calc_hr_024
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    # Headers
    ws.cell(row=1, column=1, value='Employee')
    for day in range(1, 22):
        ws.cell(row=1, column=day + 1, value=f'Day{day}')
    # Summary headers in W1, X1, Y1 (columns 23, 24, 25)
    ws.cell(row=1, column=23, value='Present')
    ws.cell(row=1, column=24, value='Absent')
    ws.cell(row=1, column=25, value='Late')

    # Alice: 17P, 2A, 2L
    alice_attendance = [
        'P', 'P', 'P', 'A', 'P', 'P', 'L', 'P', 'P', 'P',
        'P', 'P', 'A', 'P', 'P', 'P', 'L', 'P', 'P', 'P', 'P'
    ]
    # Bob: 19P, 1A, 1L
    bob_attendance = [
        'P', 'P', 'P', 'P', 'P', 'P', 'P', 'P', 'L', 'P',
        'P', 'P', 'P', 'P', 'A', 'P', 'P', 'P', 'P', 'P', 'P'
    ]
    # Carol: 15P, 4A, 2L
    carol_attendance = [
        'P', 'A', 'P', 'P', 'L', 'P', 'A', 'P', 'P', 'P',
        'A', 'P', 'P', 'L', 'P', 'P', 'A', 'P', 'P', 'P', 'P'
    ]

    # Verify counts
    assert alice_attendance.count('P') == 17
    assert alice_attendance.count('A') == 2
    assert alice_attendance.count('L') == 2
    assert bob_attendance.count('P') == 19
    assert bob_attendance.count('A') == 1
    assert bob_attendance.count('L') == 1
    assert carol_attendance.count('P') == 15
    assert carol_attendance.count('A') == 4
    assert carol_attendance.count('L') == 2

    employees = [
        ('Alice', alice_attendance),
        ('Bob', bob_attendance),
        ('Carol', carol_attendance),
    ]

    for r, (name, attendance) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=name)
        for c, status in enumerate(attendance, 2):
            ws.cell(row=r, column=c, value=status)

    # W2:Y4 are intentionally left EMPTY - the task is to add COUNTIF formulas there

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    for col_letter_idx in range(2, 23):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_letter_idx)].width = 6
    ws.column_dimensions['W'].width = 10
    ws.column_dimensions['X'].width = 10
    ws.column_dimensions['Y'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
