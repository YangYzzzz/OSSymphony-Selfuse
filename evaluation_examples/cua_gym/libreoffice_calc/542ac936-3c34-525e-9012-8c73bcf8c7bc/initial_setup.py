"""
Initial Setup: Case number sequence fill task
Task ID: osworld_calc_fill_sequence_numbers_006
Domain: libreoffice_calc

Creates a spreadsheet with case records grouped by Category (column B).
Column A (Case Number) is intentionally left empty — the agent must fill it
with per-category sequential labels like Case_001, Case_002, ... restarting
for each new category.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Cases"

    # --- Headers ---
    ws.cell(row=1, column=1, value="Case Number")
    ws.cell(row=1, column=2, value="Category")
    ws.cell(row=1, column=3, value="Description")
    ws.cell(row=1, column=4, value="Assigned To")
    ws.cell(row=1, column=5, value="Status")
    ws.cell(row=1, column=6, value="Priority")
    ws.cell(row=1, column=7, value="Date Opened")

    # --- Data: 5 categories with different row counts ---
    # Category "Network Issues"    : 6 cases
    # Category "Hardware Failure"  : 5 cases
    # Category "Software Bug"      : 4 cases
    # Category "User Access"       : 7 cases
    # Category "Data Recovery"     : 4 cases
    # Total: 26 data rows (rows 2-27)

    data = [
        # Category: Network Issues (6 cases)
        ("", "Network Issues", "VPN connection drops intermittently during peak hours", "James Morley", "Open", "High", "2025-01-06"),
        ("", "Network Issues", "Firewall blocking internal API endpoints after upgrade", "Sandra Wu", "In Progress", "Critical", "2025-01-08"),
        ("", "Network Issues", "DNS resolution failing for new domain registrations", "James Morley", "Open", "Medium", "2025-01-09"),
        ("", "Network Issues", "Wi-Fi signal unstable in conference rooms B2 and B3", "Priya Nair", "Resolved", "Low", "2025-01-12"),
        ("", "Network Issues", "Bandwidth throttling reported by remote workers on fiber plan", "Sandra Wu", "Open", "Medium", "2025-01-14"),
        ("", "Network Issues", "VLAN misconfiguration causing cross-segment traffic leak", "Priya Nair", "In Progress", "Critical", "2025-01-15"),
        # Category: Hardware Failure (5 cases)
        ("", "Hardware Failure", "Workstation WS-0042 SSD showing reallocated sectors", "Carlos Reyes", "Open", "High", "2025-01-07"),
        ("", "Hardware Failure", "Printer in Room 304 paper feed jams on envelopes", "Dana Fitzgerald", "Resolved", "Low", "2025-01-10"),
        ("", "Hardware Failure", "Server rack UPS battery test failed — replacement needed", "Carlos Reyes", "In Progress", "Critical", "2025-01-11"),
        ("", "Hardware Failure", "Monitor flickering at 144 Hz on DisplayPort cable", "Dana Fitzgerald", "Open", "Medium", "2025-01-13"),
        ("", "Hardware Failure", "Keyboard controller error on docking stations after firmware update", "Carlos Reyes", "Open", "Medium", "2025-01-16"),
        # Category: Software Bug (4 cases)
        ("", "Software Bug", "CRM module crashes on bulk export exceeding 5 000 records", "Mei-Ling Huang", "Open", "High", "2025-01-05"),
        ("", "Software Bug", "Date picker widget shows incorrect week numbers in January", "Lucas Ferreira", "In Progress", "Medium", "2025-01-07"),
        ("", "Software Bug", "Email notification template missing footer logo after v3.4 deploy", "Mei-Ling Huang", "Resolved", "Low", "2025-01-09"),
        ("", "Software Bug", "Search indexer skips documents with special characters in filename", "Lucas Ferreira", "Open", "High", "2025-01-14"),
        # Category: User Access (7 cases)
        ("", "User Access", "New hire Alicia Tran cannot log in to ERP portal", "Omar Hassan", "Open", "High", "2025-01-06"),
        ("", "User Access", "Password reset emails landing in spam for Outlook 365 users", "Rachel Kim", "In Progress", "Medium", "2025-01-08"),
        ("", "User Access", "Role 'Finance Analyst' missing read permissions on report module", "Omar Hassan", "Open", "Critical", "2025-01-09"),
        ("", "User Access", "SSO token expiry set too short — users logged out every 15 min", "Rachel Kim", "Resolved", "Medium", "2025-01-11"),
        ("", "User Access", "Shared mailbox 'support@' inaccessible after mailbox migration", "Omar Hassan", "Open", "High", "2025-01-12"),
        ("", "User Access", "Two-factor auth bypass reported on legacy admin console path", "Rachel Kim", "In Progress", "Critical", "2025-01-13"),
        ("", "User Access", "User profile photos not displaying in Teams after AD sync", "Omar Hassan", "Open", "Low", "2025-01-15"),
        # Category: Data Recovery (4 cases)
        ("", "Data Recovery", "Accidental deletion of Q4 financial reports from shared drive", "Sofia Andreeva", "Open", "Critical", "2025-01-04"),
        ("", "Data Recovery", "Backup restoration failed due to corrupt archive on tape BK-117", "Ivan Petrov", "In Progress", "High", "2025-01-08"),
        ("", "Data Recovery", "Database transaction log overflow — 3 hrs of records potentially lost", "Sofia Andreeva", "Open", "Critical", "2025-01-10"),
        ("", "Data Recovery", "Excel file recovered from recycle bin but charts are missing", "Ivan Petrov", "Resolved", "Medium", "2025-01-14"),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            # Column A (c==1) must remain None/empty — agent fills it in
            ws.cell(row=r, column=c, value=None if (c == 1) else val)

    # Column A is empty (column 1) — deliberately left blank for the task

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
