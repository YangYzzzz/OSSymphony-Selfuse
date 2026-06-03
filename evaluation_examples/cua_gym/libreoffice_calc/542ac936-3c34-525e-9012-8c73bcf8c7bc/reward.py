"""
Reward Script: Fill category-specific case number sequences in column A
Task ID: osworld_calc_fill_sequence_numbers_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): All 26 data rows in column A are filled with Case_XXX labels
  Component 2 (0.35): Sequence resets to Case_001 at the start of each category group
  Component 3 (0.25): Each category has a complete, correctly-ordered sequence with no gaps
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_fill_sequence_numbers_006'

# Expected category structure derived from the spreadsheet layout:
# Row 1: header
# Rows 2-7:  Network Issues  (6 cases)
# Rows 8-12: Hardware Failure (5 cases)
# Rows 13-16: Software Bug   (4 cases)
# Rows 17-23: User Access    (7 cases)
# Rows 24-27: Data Recovery  (4 cases)
EXPECTED_GROUPS = [
    ('Network Issues', list(range(2, 8))),    # rows 2..7  -> Case_001..Case_006
    ('Hardware Failure', list(range(8, 13))), # rows 8..12 -> Case_001..Case_005
    ('Software Bug', list(range(13, 17))),    # rows 13..16-> Case_001..Case_004
    ('User Access', list(range(17, 24))),     # rows 17..23-> Case_001..Case_007
    ('Data Recovery', list(range(24, 28))),   # rows 24..27-> Case_001..Case_004
]

CASE_PATTERN = re.compile(r'^Case_(\d{3})$', re.IGNORECASE)


def verify_task(file_path):
    """
    Verify that column A has been filled with per-category sequential case numbers
    that reset to Case_001 for each new category in column B.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # ----------------------------------------------------------------
    # Component 1: All 26 data rows in column A are filled with
    #              a Case_XXX label (0.40 points)
    # This FAILS on initial (all None) and PASSES on golden (all filled).
    # ----------------------------------------------------------------
    try:
        filled_count = 0
        total_data_rows = 26  # rows 2..27
        for row in range(2, 28):
            val = ws.cell(row=row, column=1).value
            if val is not None and CASE_PATTERN.match(str(val)):
                filled_count += 1
            else:
                print(f"FAIL Comp1: Row {row} col A = {repr(val)} (expected Case_XXX pattern)")

        if filled_count == total_data_rows:
            print(f"PASS: Component 1 — all {total_data_rows} data rows have Case_XXX labels (0.40 pts)")
            total_score += 0.40
        else:
            print(f"FAIL: Component 1 — only {filled_count}/{total_data_rows} rows have valid Case_XXX labels")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: Each category group starts fresh at Case_001 (0.35 points)
    # For each category, the first row must be Case_001.
    # ----------------------------------------------------------------
    try:
        restart_ok = sum(
            1 for category, rows in EXPECTED_GROUPS
            if ws.cell(row=rows[0], column=1).value is not None
            and str(ws.cell(row=rows[0], column=1).value).upper() == 'CASE_001'
        )
        restart_total = len(EXPECTED_GROUPS)  # 5 groups
        for category, rows in EXPECTED_GROUPS:
            first_row = rows[0]
            val = ws.cell(row=first_row, column=1).value
            if val is None or str(val).upper() != 'CASE_001':
                print(f"FAIL Comp2: Category '{category}' first row {first_row} = {repr(val)} (expected 'Case_001')")

        comp2_score = round(0.35 * restart_ok / restart_total, 4)
        if restart_ok == restart_total:
            print(f"PASS: Component 2 — all {restart_total} categories start with Case_001 (0.35 pts)")
        elif restart_ok > 0:
            print(f"PARTIAL: Component 2 — {restart_ok}/{restart_total} categories start with Case_001 ({comp2_score} pts)")
        else:
            print(f"FAIL: Component 2 — no categories start with Case_001 (0.0 pts)")
        if comp2_score > 0:
            total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Within each category, the sequence is complete and
    #              correctly ordered with no gaps (0.25 points)
    # For each group, values must be Case_001, Case_002, ..., Case_00N
    # in exactly that order.
    # ----------------------------------------------------------------
    try:
        group_ok = 0
        group_total = len(EXPECTED_GROUPS)  # 5 groups
        for category, rows in EXPECTED_GROUPS:
            mismatches = []
            for seq_idx, row in enumerate(rows, 1):
                expected_label = f'Case_{seq_idx:03d}'
                actual_val = ws.cell(row=row, column=1).value
                if actual_val is None or str(actual_val).upper() != expected_label.upper():
                    mismatches.append((row, expected_label, actual_val))
            if not mismatches:
                group_ok += 1
            else:
                for row, exp, got in mismatches:
                    print(f"FAIL Comp3: Category '{category}', row {row}: expected '{exp}', got {repr(got)}")

        comp3_score = round(0.25 * group_ok / group_total, 4)
        if group_ok == group_total:
            print(f"PASS: Component 3 — all {group_total} category sequences are complete and correctly ordered (0.25 pts)")
        elif group_ok > 0:
            print(f"PARTIAL: Component 3 — {group_ok}/{group_total} groups have correct sequences ({comp3_score} pts)")
        else:
            print(f"FAIL: Component 3 — no category has a fully correct sequence (0.0 pts)")
        if comp3_score > 0:
            total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against the canonical artifact path on this VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
