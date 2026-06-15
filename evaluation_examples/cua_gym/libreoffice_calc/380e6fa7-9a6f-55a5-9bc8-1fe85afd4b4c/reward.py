"""
Reward Script: Weekly Project Status Report Template
Task ID: calc_gen_template_036
Domain: libreoffice_calc
Scoring:
  Component 1: Key formulas present (C2=TODAY, D12=burn rate, D14=utilization) — 0.25 pts
  Component 2: RAG status formula in C18 with correct logic — 0.25 pts
  Component 3: Data validation dropdown on E5:E10 (On Track/At Risk/Delayed) — 0.20 pts
  Component 4: Conditional formatting on C18 with red/amber/green fills — 0.15 pts
  Component 5: Sheet protection enabled with input cells unlocked — 0.15 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_template_036'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'WeeklyStatus' not in wb.sheetnames:
        print("FAIL: Sheet 'WeeklyStatus' not found")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['WeeklyStatus']

    # Component 1: Key formulas present — C2 (TODAY), D12 (burn rate), D14 (utilization) (0.25 points)
    # These cells are None in the initial file and must contain formulas in the golden file.
    try:
        c2_val = ws['C2'].value
        d12_val = ws['D12'].value
        d14_val = ws['D14'].value

        c2_ok = (c2_val is not None and isinstance(c2_val, str) and
                 'TODAY' in c2_val.upper())
        d12_ok = (d12_val is not None and isinstance(d12_val, str) and
                  'C12' in d12_val.upper() and 'B12' in d12_val.upper())
        d14_ok = (d14_val is not None and isinstance(d14_val, str) and
                  'C14' in d14_val.upper() and 'B14' in d14_val.upper())

        formulas_present = sum([c2_ok, d12_ok, d14_ok])

        if formulas_present == 3:
            print(f"PASS: Component 1 — All 3 key formulas present: C2={repr(c2_val)}, D12={repr(d12_val)}, D14={repr(d14_val)} (0.25 pts)")
            total_score += 0.25
        elif formulas_present >= 1:
            partial = round(0.25 * formulas_present / 3, 4)
            print(f"PARTIAL: Component 1 — {formulas_present}/3 formulas present (c2={c2_ok}, d12={d12_ok}, d14={d14_ok}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No key formulas found: C2={repr(c2_val)}, D12={repr(d12_val)}, D14={repr(d14_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: RAG status formula in C18 (0.25 points)
    # C18 must contain an IF/COUNTIF formula referencing E5:E10, "Delayed", "At Risk", and D12
    # The initial file has None in C18.
    try:
        c18_val = ws['C18'].value

        c18_has_formula = (c18_val is not None and isinstance(c18_val, str)
                           and c18_val.strip().startswith('='))
        c18_has_delayed = c18_has_formula and 'DELAYED' in c18_val.upper()
        c18_has_at_risk = c18_has_formula and 'AT RISK' in c18_val.upper()
        c18_has_rag_colors = c18_has_formula and ('RED' in c18_val.upper() or
                                                   'AMBER' in c18_val.upper() or
                                                   'GREEN' in c18_val.upper())
        c18_references_milestones = c18_has_formula and 'E5' in c18_val.upper()
        c18_references_budget = c18_has_formula and 'D12' in c18_val.upper()

        rag_criteria = [c18_has_formula, c18_has_delayed, c18_has_at_risk,
                        c18_has_rag_colors, c18_references_milestones, c18_references_budget]
        rag_score = sum(rag_criteria)

        if rag_score == 6:
            print(f"PASS: Component 2 — RAG formula in C18 with all required logic (0.25 pts): {repr(c18_val[:80])}")
            total_score += 0.25
        elif rag_score >= 3:
            partial = round(0.25 * rag_score / 6, 4)
            print(f"PARTIAL: Component 2 — RAG formula partially correct ({rag_score}/6 criteria): "
                  f"has_formula={c18_has_formula}, has_delayed={c18_has_delayed}, "
                  f"has_at_risk={c18_has_at_risk}, rag_colors={c18_has_rag_colors}, "
                  f"ref_milestones={c18_references_milestones}, ref_budget={c18_references_budget} ({partial} pts)")
            total_score += partial
        elif rag_score >= 1:
            partial = round(0.25 * rag_score / 6, 4)
            print(f"PARTIAL: Component 2 — Minimal RAG formula ({rag_score}/6 criteria) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — C18 is empty or not a formula: {repr(c18_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data validation dropdown on E5:E10 with On Track/At Risk/Delayed (0.20 points)
    # The initial file has no data validations at all.
    try:
        validations = ws.data_validations.dataValidation
        found_milestone_dv = False
        dv_formula = None
        dv_sqref = None

        for dv in validations:
            if dv.type == 'list':
                formula = dv.formula1 or ''
                formula_clean = formula.strip('"').replace('"', '')
                # Check if it contains the three milestone statuses
                has_on_track = 'On Track' in formula_clean or 'ON TRACK' in formula_clean.upper()
                has_at_risk = 'At Risk' in formula_clean or 'AT RISK' in formula_clean.upper()
                has_delayed = 'Delayed' in formula_clean or 'DELAYED' in formula_clean.upper()

                sqref_str = str(dv.sqref)
                covers_e5_e10 = 'E5' in sqref_str or 'E5:E10' in sqref_str

                if (has_on_track or has_at_risk or has_delayed) and covers_e5_e10:
                    found_milestone_dv = True
                    dv_formula = formula
                    dv_sqref = sqref_str
                    all_three = has_on_track and has_at_risk and has_delayed
                    if all_three:
                        print(f"PASS: Component 3 — Dropdown validation on {sqref_str} with all 3 statuses: {repr(formula)} (0.20 pts)")
                        total_score += 0.20
                    else:
                        print(f"PARTIAL: Component 3 — Dropdown validation on {sqref_str} found but missing some statuses: {repr(formula)} (0.10 pts)")
                        total_score += 0.10
                    break

        if not found_milestone_dv:
            if len(validations) == 0:
                print(f"FAIL: Component 3 — No data validations found in the file")
            else:
                print(f"FAIL: Component 3 — {len(validations)} validation(s) found but none match E5:E10 milestone dropdown")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on C18 with red/amber/green fills (0.15 points)
    # The initial file has no conditional formatting.
    try:
        cf_rules_count = 0
        has_red_cf = False
        has_amber_cf = False
        has_green_cf = False

        for cf_range in ws.conditional_formatting:
            range_str = str(cf_range)
            if 'C18' in range_str or range_str == 'C18':
                for rule in ws.conditional_formatting[cf_range]:
                    cf_rules_count += 1
                    # Check fill colors
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fg_rgb = rule.dxf.fill.fgColor.rgb
                            if fg_rgb:
                                fg_upper = fg_rgb.upper()
                                # Red: FFFF0000
                                if 'FF0000' in fg_upper:
                                    has_red_cf = True
                                # Amber/Yellow: FFFFC000 or similar
                                elif 'FFC0' in fg_upper or 'FFFF00' in fg_upper or 'FFA5' in fg_upper:
                                    has_amber_cf = True
                                # Green: FF00B050 or FF00FF00 or similar
                                elif '00B050' in fg_upper or '00FF00' in fg_upper or '92D050' in fg_upper:
                                    has_green_cf = True
                        except Exception:
                            pass
                    # Also check via formula content for rule color mapping
                    formula_list = getattr(rule, 'formula', []) or []
                    for f in formula_list:
                        f_upper = str(f).upper()
                        if 'RED' in f_upper:
                            pass  # color already determined by fill
                        if 'AMBER' in f_upper:
                            pass
                        if 'GREEN' in f_upper:
                            pass

        color_count = sum([has_red_cf, has_amber_cf, has_green_cf])

        if cf_rules_count >= 3 and color_count == 3:
            print(f"PASS: Component 4 — Conditional formatting on C18: {cf_rules_count} rules, red={has_red_cf}, amber={has_amber_cf}, green={has_green_cf} (0.15 pts)")
            total_score += 0.15
        elif cf_rules_count >= 1 and color_count >= 1:
            partial = round(0.15 * color_count / 3, 4)
            print(f"PARTIAL: Component 4 — CF on C18 partial: {cf_rules_count} rules, {color_count}/3 colors ({partial} pts)")
            total_score += partial
        elif cf_rules_count >= 1:
            print(f"PARTIAL: Component 4 — CF rules on C18 found ({cf_rules_count}) but colors not identified (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on C18")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Sheet protection enabled, with input cells (C1, C12, C14, C16) unlocked (0.15 points)
    # Initial file has no sheet protection and all cells are locked=True.
    try:
        protection_enabled = ws.protection.sheet

        if not protection_enabled:
            print(f"FAIL: Component 5 — Sheet protection is not enabled")
        else:
            # Check that input cells are unlocked (locked=False)
            input_cells = ['C1', 'C12', 'C14', 'C16']
            unlocked = []
            still_locked = []
            for c in input_cells:
                cell = ws[c]
                if cell.protection.locked is False:
                    unlocked.append(c)
                else:
                    still_locked.append(c)

            if len(unlocked) == 4:
                print(f"PASS: Component 5 — Sheet protection enabled; all input cells unlocked: {unlocked} (0.15 pts)")
                total_score += 0.15
            elif len(unlocked) >= 2:
                partial = round(0.15 * len(unlocked) / 4, 4)
                print(f"PARTIAL: Component 5 — Sheet protected; {len(unlocked)}/4 input cells unlocked: {unlocked}, still locked: {still_locked} ({partial} pts)")
                total_score += partial
            elif protection_enabled:
                print(f"PARTIAL: Component 5 — Sheet protection enabled but input cells not unlocked (0.05 pts)")
                total_score += 0.05
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
