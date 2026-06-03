"""
Initial Setup: Weekly Project Status Report Template
Task ID: calc_gen_template_036
Domain: libreoffice_calc

Creates the pre-task initial state: a WeeklyStatus sheet with labels
at EXACT cell positions per task spec, but WITHOUT task-completion elements:
- C2: empty (no TODAY() formula — that's the task)
- E5:E10: no dropdown validation (that's the task)
- D12: empty (no burn rate formula — that's the task)
- D14: empty (no utilization formula — that's the task)
- C18: empty (no RAG status formula — that's the task)
- No conditional formatting on C18 (that's the task)
- No sheet protection (that's the task)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_template_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'WeeklyStatus'

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # ===========================================================
    # ROWS 1-2: Project Info Section
    # Exact positions from task spec:
    #   B1 = 'Project Name', C1 = input cell (empty)
    #   B2 = 'Report Date',  C2 = will be TODAY() — leave empty for now
    # ===========================================================
    ws['B1'] = 'Project Name'
    ws['B1'].font = Font(bold=True)
    # C1: editable input cell — left empty intentionally

    ws['B2'] = 'Report Date'
    ws['B2'].font = Font(bold=True)
    # C2: will be =TODAY() in golden — left empty here (MUST NOT have formula)

    # Row 3: Section divider / Milestones label
    ws['B3'] = 'MILESTONES'
    ws['B3'].font = Font(bold=True, size=11)
    ws['B3'].fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    ws.merge_cells('B3:F3')
    ws['B3'].alignment = Alignment(horizontal='left')

    # ===========================================================
    # ROWS 4-10: Milestones Section
    # Task spec: "Milestones section (rows 4-10): Status column with dropdown"
    # Row 4 = column headers, rows 5-10 = 6 milestone data rows
    # Status column (E) left empty — no dropdown validation in initial
    # ===========================================================
    ws['B4'] = 'Milestone'
    ws['C4'] = 'Owner'
    ws['D4'] = 'Due Date'
    ws['E4'] = 'Status'
    ws['F4'] = 'Notes'
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws[f'{col}4'].font = Font(bold=True)
        ws[f'{col}4'].fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')

    # Rows 5-10: 6 milestone data rows (Status column E left empty — NO dropdown)
    milestones = [
        ('Requirements Finalization', 'Sarah Chen', '2025-01-15'),
        ('Architecture Design Review', 'Marcus Johnson', '2025-02-01'),
        ('Backend API Development', 'Priya Patel', '2025-02-28'),
        ('Frontend UI Implementation', 'David Kim', '2025-03-15'),
        ('Integration Testing', 'Emma Rodriguez', '2025-03-30'),
        ('Production Deployment', 'Aisha Okonkwo', '2025-04-25'),
    ]
    for i, (milestone, owner, due_date) in enumerate(milestones, 5):
        ws.cell(row=i, column=2, value=milestone)   # B: milestone name
        ws.cell(row=i, column=3, value=owner)        # C: owner
        ws.cell(row=i, column=4, value=due_date)     # D: due date
        # E: Status — NO value, NO dropdown validation in initial
        ws.cell(row=i, column=6, value='')           # F: notes (empty)

    # ===========================================================
    # ROW 11: Budget Section Header
    # ROW 12: Budget data row
    # Task spec: "Budget (B12), Spent (C12 — input), Burn Rate (D12: =C12/B12 %)"
    # B12 = budget value number, C12 = spent input (empty), D12 = no formula yet
    # ===========================================================
    ws['B11'] = 'BUDGET'
    ws['B11'].font = Font(bold=True, size=11)
    ws['B11'].fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    ws.merge_cells('B11:F11')
    ws['B11'].alignment = Alignment(horizontal='left')

    ws['B12'] = 500000    # Budget value (editable reference value)
    ws['C12'] = None      # Spent ($) — input cell, left empty (MUST NOT have value)
    ws['D12'] = None      # Burn Rate — NO formula yet (MUST NOT have =C12/B12)

    ws['B12'].number_format = '$#,##0.00'

    # ===========================================================
    # ROW 13: Team Utilization Section Header
    # ROW 14: Team data row
    # Task spec: "Capacity Hours (B14), Used Hours (C14 — input), Utilization (D14: =C14/B14 %)"
    # B14 = capacity hours, C14 = used hours input (empty), D14 = no formula yet
    # ===========================================================
    ws['B13'] = 'TEAM UTILIZATION'
    ws['B13'].font = Font(bold=True, size=11)
    ws['B13'].fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    ws.merge_cells('B13:F13')
    ws['B13'].alignment = Alignment(horizontal='left')

    ws['B14'] = 480       # Capacity hours
    ws['C14'] = None      # Used hours — input cell, left empty (MUST NOT have value)
    ws['D14'] = None      # Utilization — NO formula yet (MUST NOT have =C14/B14)

    # ===========================================================
    # ROW 15: Open Issues Section Header
    # ROW 16: Issues data
    # Task spec: "Open Issues count (C16 — input)"
    # C16 = open issues count input (empty)
    # ===========================================================
    ws['B15'] = 'OPEN ISSUES'
    ws['B15'].font = Font(bold=True, size=11)
    ws['B15'].fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    ws.merge_cells('B15:F15')
    ws['B15'].alignment = Alignment(horizontal='left')

    ws['B16'] = 'Open Issues Count'
    ws['B16'].font = Font(bold=True)
    ws['C16'] = None      # Input cell, left empty (MUST NOT have value)

    # ===========================================================
    # ROW 17: Overall RAG Status Section Header
    # ROW 18: RAG Status data
    # Task spec: "Overall RAG status (C18): formula — Red/Amber/Green"
    # C18 = RAG formula — NO formula in initial (MUST be empty)
    # ===========================================================
    ws['B17'] = 'OVERALL STATUS'
    ws['B17'].font = Font(bold=True, size=11)
    ws['B17'].fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    ws.merge_cells('B17:F17')
    ws['B17'].alignment = Alignment(horizontal='left')

    ws['B18'] = 'RAG Status'
    ws['B18'].font = Font(bold=True)
    ws['C18'] = None      # RAG formula — NO formula yet, NO conditional formatting

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: WeeklyStatus')
    print('Key cells:')
    print('  B1=Project Name label, C1=empty input')
    print('  B2=Report Date label, C2=empty (no TODAY formula)')
    print('  Rows 5-10: milestone data, E col empty (no dropdown)')
    print('  B12=500000 budget, C12=empty, D12=empty (no formula)')
    print('  B14=480 capacity, C14=empty, D14=empty (no formula)')
    print('  B16=label, C16=empty input')
    print('  B18=RAG Status label, C18=empty (no formula, no CF)')


create_initial()
