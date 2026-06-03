"""
Initial Setup: Bank reconciliation spreadsheet
Task ID: calc_fin_bank_reconciliation_034
Domain: libreoffice_calc

Creates a BankRec sheet with:
- Bank statement ending balance in B1
- Book balance per GL in B8
- Labeled rows for deposits in transit (B3:B6) and outstanding checks (B9:B14)
- Labeled summary rows for totals and adjusted balances (formulas NOT included)
- Bank charges (B18), interest earned (B19) placeholders
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_bank_reconciliation_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BankRec'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 18

    # --- Title ---
    ws['A1'] = 'Bank Statement Ending Balance'
    ws['B1'] = 125450.00
    ws['B1'].number_format = '#,##0.00'

    # --- Deposits in Transit section ---
    ws['A2'] = 'Add: Deposits in Transit'
    ws['A2'].font = Font(bold=True)

    ws['A3'] = 'Deposit 1 (Mar 28)'
    ws['A4'] = 'Deposit 2 (Mar 29)'
    ws['A5'] = 'Deposit 3 (Mar 30)'
    ws['A6'] = 'Deposit 4 (Mar 31)'
    # B3:B6 intentionally left empty (user-entered)

    ws['A7'] = 'Total Deposits in Transit'
    # B7 intentionally left empty (formula will be added in golden)

    # --- Book Balance section ---
    ws['A8'] = 'Book Balance per GL'
    ws['B8'] = 115200.00
    ws['B8'].number_format = '#,##0.00'

    # --- Outstanding Checks section ---
    ws['A9'] = 'Less: Outstanding Checks'
    ws['A9'].font = Font(bold=True)

    ws['A10'] = 'Check #1042 - Office Supplies Co.'
    ws['A11'] = 'Check #1043 - City Utilities'
    ws['A12'] = 'Check #1044 - Landlord Properties LLC'
    ws['A13'] = 'Check #1045 - Tech Solutions Inc.'
    ws['A14'] = 'Check #1046 - Insurance Corp.'
    # Note: row 9 in task context is actually the first check row
    # The section header in A9 serves as label; B9:B14 for check amounts
    # Adjusting to match context: B9:B14 are individual check amounts
    # B9:B14 intentionally left empty (user-entered)

    ws['A15'] = 'Total Outstanding Checks'
    # B15 intentionally left empty (formula will be added in golden)

    ws['A16'] = 'Adjusted Bank Balance'
    # B16 intentionally left empty (formula + formatting will be added in golden)

    # --- Bank Book Side section ---
    ws['A17'] = 'Book Balance Adjustments'
    ws['A17'].font = Font(bold=True)

    ws['A18'] = 'Less: Bank Charges'
    # B18 intentionally left empty (user-entered)

    ws['A19'] = 'Add: Interest Earned'
    # B19 intentionally left empty (user-entered)

    ws['A20'] = 'Adjusted Book Balance'
    # B20 intentionally left empty (formula + formatting will be added in golden)

    ws['A21'] = ''

    ws['A22'] = 'Difference (Bank - Book)'
    # B22 intentionally left empty (formula + conditional formatting will be added in golden)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
