"""
Reward Script: Set up monthly bank reconciliation spreadsheet
Task ID: calc_fin_bank_reconciliation_034
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Bank-side totals — SUM formulas in B7 (deposits) and B15 (checks)   (0.25 pts)
  Component 2: Adjusted Bank Balance formula in B16 (=B1+B7-B15)                   (0.20 pts)
  Component 3: Book-side formula in B20 (=B8-B18+B19)                              (0.15 pts)
  Component 4: B16 and B20 have bold font and currency number format                (0.15 pts)
  Component 5: Difference formula in B22 (=B16-B20) with currency format           (0.10 pts)
  Component 6: Conditional formatting on B22 (red != 0, green == 0)                (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fin_bank_reconciliation_034'


def normalize_formula(formula):
    """Normalize a formula string for comparison: uppercase, strip spaces and = prefix."""
    if not isinstance(formula, str):
        return ''
    return formula.strip().upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: BankRec sheet must exist
    if 'BankRec' not in wb.sheetnames:
        print("CRITICAL: Sheet 'BankRec' not found. Cannot proceed.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['BankRec']

    # Component 1: Bank-side SUM formulas — B7 and B15 (0.25 points)
    # B7 should contain =SUM(B3:B6) for Total Deposits in Transit
    # B15 should contain =SUM(B9:B14) for Total Outstanding Checks
    # Both are ABSENT in the initial file (all None)
    try:
        b7_val = ws['B7'].value
        b15_val = ws['B15'].value

        b7_ok = (isinstance(b7_val, str) and
                 'SUM' in b7_val.upper() and
                 'B3' in b7_val.upper() and
                 'B6' in b7_val.upper())
        b15_ok = (isinstance(b15_val, str) and
                  'SUM' in b15_val.upper() and
                  'B9' in b15_val.upper() and
                  'B14' in b15_val.upper())

        if b7_ok and b15_ok:
            print(f"PASS: Component 1 — B7={repr(b7_val)}, B15={repr(b15_val)} (0.25 pts)")
            total_score += 0.25
        elif b7_ok:
            print(f"PASS (partial): Component 1 — B7 SUM found but B15={repr(b15_val)} missing/wrong (0.10 pts)")
            total_score += 0.10
        elif b15_ok:
            print(f"PASS (partial): Component 1 — B15 SUM found but B7={repr(b7_val)} missing/wrong (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — B7={repr(b7_val)} (expected SUM(B3:B6)), B15={repr(b15_val)} (expected SUM(B9:B14))")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Adjusted Bank Balance formula in B16 (0.20 points)
    # B16 should contain =B1+B7-B15 (bank statement balance + deposits - checks)
    # ABSENT in initial file
    try:
        b16_val = ws['B16'].value
        b16_ok = (isinstance(b16_val, str) and
                  'B1' in b16_val.upper() and
                  'B7' in b16_val.upper() and
                  'B15' in b16_val.upper())
        if b16_ok:
            print(f"PASS: Component 2 — B16={repr(b16_val)} adjusted bank balance formula (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — B16={repr(b16_val)}, expected formula referencing B1, B7, B15")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Book-side Adjusted Book Balance formula in B20 (0.15 points)
    # B20 should contain =B8-B18+B19 (book balance - bank charges + interest)
    # ABSENT in initial file
    try:
        b20_val = ws['B20'].value
        b20_ok = (isinstance(b20_val, str) and
                  'B8' in b20_val.upper() and
                  'B18' in b20_val.upper() and
                  'B19' in b20_val.upper())
        if b20_ok:
            print(f"PASS: Component 3 — B20={repr(b20_val)} adjusted book balance formula (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — B20={repr(b20_val)}, expected formula referencing B8, B18, B19")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: B16 and B20 have bold font and currency number format (0.15 points)
    # Both ABSENT in initial file (no bold, no currency format)
    try:
        b16_cell = ws['B16']
        b20_cell = ws['B20']

        b16_bold = b16_cell.font.bold is True
        b20_bold = b20_cell.font.bold is True
        b16_currency = '$' in b16_cell.number_format or '#,##0' in b16_cell.number_format
        b20_currency = '$' in b20_cell.number_format or '#,##0' in b20_cell.number_format

        bold_ok = b16_bold and b20_bold
        currency_ok = b16_currency and b20_currency

        if bold_ok and currency_ok:
            print(f"PASS: Component 4 — B16 bold={b16_bold}, fmt={repr(b16_cell.number_format)}; "
                  f"B20 bold={b20_bold}, fmt={repr(b20_cell.number_format)} (0.15 pts)")
            total_score += 0.15
        elif bold_ok or currency_ok:
            print(f"PASS (partial): Component 4 — bold_ok={bold_ok}, currency_ok={currency_ok} (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 4 — B16 bold={b16_bold}, fmt={repr(b16_cell.number_format)}; "
                  f"B20 bold={b20_bold}, fmt={repr(b20_cell.number_format)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Difference formula in B22 with currency format (0.10 points)
    # B22 should contain =B16-B20 with currency number format
    # ABSENT in initial file
    try:
        b22_cell = ws['B22']
        b22_val = b22_cell.value
        b22_formula_ok = (isinstance(b22_val, str) and
                          'B16' in b22_val.upper() and
                          'B20' in b22_val.upper())
        b22_currency_ok = '$' in b22_cell.number_format or '#,##0' in b22_cell.number_format

        if b22_formula_ok and b22_currency_ok:
            print(f"PASS: Component 5 — B22={repr(b22_val)}, fmt={repr(b22_cell.number_format)} (0.10 pts)")
            total_score += 0.10
        elif b22_formula_ok:
            print(f"PASS (partial): Component 5 — B22 formula found but no currency format; fmt={repr(b22_cell.number_format)} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — B22={repr(b22_val)}, expected formula referencing B16 and B20")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on B22 — red for non-zero, green for zero (0.15 points)
    # NO conditional formatting in initial file
    try:
        cf = ws.conditional_formatting
        b22_cf_count = sum(1 for cf_obj in cf if 'B22' in str(cf_obj))
        red_rule_count = 0
        green_rule_count = 0

        for cf_obj in cf:
            if 'B22' not in str(cf_obj):
                continue
            # Access rules via .rules attribute (not by iterating cf_obj directly)
            rules = getattr(cf_obj, 'rules', [])
            for rule in rules:
                operator = getattr(rule, 'operator', '')
                if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                    fill_color = None
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb
                    except Exception:
                        pass
                    if fill_color and 'FF0000' in fill_color.upper() and operator == 'notEqual':
                        red_rule_count += 1
                    if fill_color and '00FF00' in fill_color.upper() and operator == 'equal':
                        green_rule_count += 1

        if b22_cf_count > 0 and red_rule_count > 0 and green_rule_count > 0:
            print(f"PASS: Component 6 — B22 has conditional formatting with red (!=0) and green (=0) rules (0.15 pts)")
            total_score += 0.15
        elif b22_cf_count > 0 and (red_rule_count > 0 or green_rule_count > 0):
            print(f"PASS (partial): Component 6 — B22 has CF, red_rules={red_rule_count}, green_rules={green_rule_count} (0.08 pts)")
            total_score += 0.08
        elif b22_cf_count > 0:
            print(f"PASS (partial): Component 6 — B22 has CF rules but no recognizable red/green fills (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No conditional formatting found for B22")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
