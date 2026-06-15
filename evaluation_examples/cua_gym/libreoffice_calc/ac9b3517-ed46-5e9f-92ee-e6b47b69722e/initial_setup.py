"""
Initial Setup: Create a workbook with 12 monthly sheets + Summary, unformatted.
Task ID: calc_gsi_044
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

MONTHS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

# Realistic employee names and departments
EMPLOYEES = [
    ('Sarah Chen', 'Engineering'),
    ('Marcus Johnson', 'Marketing'),
    ('Priya Patel', 'Sales'),
    ('David Kim', 'Operations'),
    ('Elena Rodriguez', 'Finance'),
    ('James Wright', 'Engineering'),
    ('Aisha Mohammed', 'Marketing'),
    ('Robert Taylor', 'Sales'),
    ('Mei-Ling Wu', 'Operations'),
    ('Carlos Fernandez', 'Finance'),
    ('Anna Kowalski', 'Engineering'),
    ('Thomas Brown', 'Sales'),
]

# Base revenue figures per employee (will vary by month)
BASE_REVENUE = [8500, 7200, 9100, 6800, 7500, 8200, 6900, 9400, 7100, 8800, 7600, 8100]
BASE_EXPENSES = [3200, 2800, 3500, 2600, 2900, 3100, 2700, 3600, 2800, 3400, 2900, 3000]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    for i, month in enumerate(MONTHS):
        if i == 0:
            ws = wb.active
            ws.title = month
        else:
            ws = wb.create_sheet(month)

        # Headers
        headers = ['Employee', 'Department', 'Revenue', 'Expenses', 'Net Profit']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        # Data rows - vary by month index for realism
        for r, (name, dept) in enumerate(EMPLOYEES, 2):
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=dept)
            # Vary revenue and expenses by month
            rev = BASE_REVENUE[r - 2] + (i * 150) + ((r - 2) * 50)
            exp = BASE_EXPENSES[r - 2] + (i * 80) + ((r - 2) * 30)
            ws.cell(row=r, column=3, value=rev)
            ws.cell(row=r, column=4, value=exp)
            ws.cell(row=r, column=5, value=rev - exp)

    # Summary sheet
    ws_summary = wb.create_sheet('Summary')
    ws_summary.cell(row=1, column=1, value='Month')
    ws_summary.cell(row=1, column=2, value='Total Revenue')
    ws_summary.cell(row=1, column=3, value='Total Expenses')
    ws_summary.cell(row=1, column=4, value='Total Net Profit')

    for idx, month in enumerate(MONTHS):
        row = idx + 2
        ws_summary.cell(row=row, column=1, value=month)
        # Calculate totals from the monthly data
        month_ws = wb[month]
        total_rev = sum(month_ws.cell(row=r, column=3).value for r in range(2, 14))
        total_exp = sum(month_ws.cell(row=r, column=4).value for r in range(2, 14))
        ws_summary.cell(row=row, column=2, value=total_rev)
        ws_summary.cell(row=row, column=3, value=total_exp)
        ws_summary.cell(row=row, column=4, value=total_rev - total_exp)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
