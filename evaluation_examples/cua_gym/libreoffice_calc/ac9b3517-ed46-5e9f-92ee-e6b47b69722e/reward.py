"""
Reward Script: Uniform formatting across monthly sheets via sheet grouping
Task ID: calc_gsi_044
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Column widths set on all 12 monthly sheets (A=22, B=18, C/D/E=15)
  Component 2 (0.30): Header background color FF4472C4 on all 12 monthly sheets
  Component 3 (0.20): Header font bold=True and size=12 on all 12 monthly sheets
  Component 4 (0.15): Header font color white (FFFFFF) on all 12 monthly sheets
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_044'

MONTHLY_SHEETS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

# Expected column widths in golden state
EXPECTED_WIDTHS = {'A': 22.0, 'B': 18.0, 'C': 15.0, 'D': 15.0, 'E': 15.0}

# Expected header background color (ARGB)
EXPECTED_FILL_RGB = 'FF4472C4'

# Expected header font size
EXPECTED_FONT_SIZE = 12.0

# Expected header font color (white, ARGB)
EXPECTED_FONT_COLOR = 'FFFFFF'  # We'll match last 6 chars of ARGB


def persist_app_state(domain):
    """Try to save any open LibreOffice document before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: all 12 monthly sheets must exist
    for name in MONTHLY_SHEETS:
        if name not in wb.sheetnames:
            print(f"CRITICAL: Monthly sheet '{name}' missing from workbook")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Column widths set on all 12 monthly sheets (0.35 points)
    # In initial state, no custom widths are set. In golden, A=22, B=18, C/D/E=15.
    try:
        sheets_with_correct_widths = 0
        for name in MONTHLY_SHEETS:
            ws = wb[name]
            all_correct = True
            for col_letter, expected_w in EXPECTED_WIDTHS.items():
                dim = ws.column_dimensions.get(col_letter)
                if dim is None or dim.width is None:
                    all_correct = False
                    break
                # Allow 1.0 tolerance for column width
                if abs(dim.width - expected_w) > 1.5:
                    all_correct = False
                    break
            if all_correct:
                sheets_with_correct_widths += 1

        if sheets_with_correct_widths == 12:
            print(f"PASS: Component 1 -- Column widths correct on all 12 monthly sheets (0.35 pts)")
            total_score += 0.35
        elif sheets_with_correct_widths >= 6:
            partial = 0.35 * (sheets_with_correct_widths / 12.0)
            print(f"PARTIAL: Component 1 -- Column widths correct on {sheets_with_correct_widths}/12 sheets ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Column widths correct on only {sheets_with_correct_widths}/12 sheets")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Header background color FF4472C4 on all 12 monthly sheets (0.30 points)
    # In initial state, fill is 00000000 (no fill). In golden, FF4472C4.
    try:
        sheets_with_correct_fill = 0
        for name in MONTHLY_SHEETS:
            ws = wb[name]
            all_filled = True
            for c in range(1, 6):  # columns A-E
                cell = ws.cell(row=1, column=c)
                try:
                    fill_rgb = cell.fill.fgColor.rgb
                except:
                    fill_rgb = None
                if fill_rgb != EXPECTED_FILL_RGB:
                    all_filled = False
                    break
            if all_filled:
                sheets_with_correct_fill += 1

        if sheets_with_correct_fill == 12:
            print(f"PASS: Component 2 -- Header background color correct on all 12 monthly sheets (0.30 pts)")
            total_score += 0.30
        elif sheets_with_correct_fill >= 6:
            partial = 0.30 * (sheets_with_correct_fill / 12.0)
            print(f"PARTIAL: Component 2 -- Header fill correct on {sheets_with_correct_fill}/12 sheets ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Header fill correct on only {sheets_with_correct_fill}/12 sheets")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Header font bold=True and size=12 on all 12 monthly sheets (0.20 points)
    # In initial state, bold=False and size=11. In golden, bold=True and size=12.
    try:
        sheets_with_correct_font = 0
        for name in MONTHLY_SHEETS:
            ws = wb[name]
            all_correct = True
            for c in range(1, 6):
                cell = ws.cell(row=1, column=c)
                if not cell.font.bold:
                    all_correct = False
                    break
                if cell.font.size is None or abs(cell.font.size - EXPECTED_FONT_SIZE) > 0.5:
                    all_correct = False
                    break
            if all_correct:
                sheets_with_correct_font += 1

        if sheets_with_correct_font == 12:
            print(f"PASS: Component 3 -- Header font (bold+size 12) correct on all 12 monthly sheets (0.20 pts)")
            total_score += 0.20
        elif sheets_with_correct_font >= 6:
            partial = 0.20 * (sheets_with_correct_font / 12.0)
            print(f"PARTIAL: Component 3 -- Header font correct on {sheets_with_correct_font}/12 sheets ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Header font correct on only {sheets_with_correct_font}/12 sheets")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Header font color white on all 12 monthly sheets (0.15 points)
    # In initial state, font color is default (dark). In golden, it's white (FFFFFF).
    try:
        sheets_with_correct_font_color = 0
        for name in MONTHLY_SHEETS:
            ws = wb[name]
            all_correct = True
            for c in range(1, 6):
                cell = ws.cell(row=1, column=c)
                try:
                    font_rgb = cell.font.color.rgb
                    # Check if the last 6 chars are FFFFFF (white)
                    if font_rgb is None or font_rgb[-6:] != EXPECTED_FONT_COLOR:
                        all_correct = False
                        break
                except:
                    all_correct = False
                    break
            if all_correct:
                sheets_with_correct_font_color += 1

        if sheets_with_correct_font_color == 12:
            print(f"PASS: Component 4 -- Header font color white on all 12 monthly sheets (0.15 pts)")
            total_score += 0.15
        elif sheets_with_correct_font_color >= 6:
            partial = 0.15 * (sheets_with_correct_font_color / 12.0)
            print(f"PARTIAL: Component 4 -- Header font color correct on {sheets_with_correct_font_color}/12 sheets ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- Header font color correct on only {sheets_with_correct_font_color}/12 sheets")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
