"""
Reward Script: Sales compensation plan modeler
Task ID: calc_sales_067
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Earned Variable formulas in row 5 (B5:F5 = Variable * Attainment Factor)
  Component 2 (0.30): Accelerator formulas in row 6 (B6:F6 = IF(attainment > 1, earned - variable, 0))
  Component 3 (0.35): Total Comp formulas in row 7 (B7:F7 = Base + Earned Variable)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_067'


def persist_app_state(domain: str):
    """Send Ctrl+S to save any unsaved changes in LibreOffice."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def check_formula(ws, coord, expected_formula):
    """Check if cell contains expected formula (case-insensitive, whitespace-normalized)."""
    val = ws[coord].value
    if not isinstance(val, str):
        return False
    return val.upper().replace(" ", "") == expected_formula.upper().replace(" ", "")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify CompModel sheet exists
    if 'CompModel' not in wb.sheetnames:
        print("CRITICAL: 'CompModel' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CompModel']

    # Column letters for the 5 scenarios
    cols = ['B', 'C', 'D', 'E', 'F']

    # Component 1: Earned Variable formulas in row 5 (0.35 points)
    # Expected: each cell = Variable(row3) * AttainmentFactor(row4)
    # e.g. B5 = =B3*B4, C5 = =C3*C4, etc.
    try:
        row5_pass = 0
        for col in cols:
            expected = f"={col}3*{col}4"
            cell_coord = f"{col}5"
            if check_formula(ws, cell_coord, expected):
                row5_pass += 1
                print(f"PASS: {cell_coord} has correct formula {expected}")
            else:
                actual = ws[cell_coord].value
                print(f"FAIL: {cell_coord} expected '{expected}', found '{actual}'")
        if row5_pass == 5:
            print(f"PASS: Component 1 — All Earned Variable formulas correct (0.35 pts)")
            total_score += 0.35
        elif row5_pass > 0:
            partial = round(0.35 * row5_pass / 5, 4)
            print(f"PARTIAL: Component 1 — {row5_pass}/5 Earned Variable formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No Earned Variable formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Accelerator formulas in row 6 (0.30 points)
    # Expected: each cell = IF(AttainmentFactor > 1, EarnedVariable - Variable, 0)
    # e.g. B6 = =IF(B4>1,B5-B3,0)
    try:
        row6_pass = 0
        for col in cols:
            expected = f"=IF({col}4>1,{col}5-{col}3,0)"
            cell_coord = f"{col}6"
            if check_formula(ws, cell_coord, expected):
                row6_pass += 1
                print(f"PASS: {cell_coord} has correct formula {expected}")
            else:
                actual = ws[cell_coord].value
                print(f"FAIL: {cell_coord} expected '{expected}', found '{actual}'")
        if row6_pass == 5:
            print(f"PASS: Component 2 — All Accelerator formulas correct (0.30 pts)")
            total_score += 0.30
        elif row6_pass > 0:
            partial = round(0.30 * row6_pass / 5, 4)
            print(f"PARTIAL: Component 2 — {row6_pass}/5 Accelerator formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No Accelerator formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Total Comp formulas in row 7 (0.35 points)
    # Expected: each cell = Base(row2) + EarnedVariable(row5)
    # e.g. B7 = =B2+B5
    try:
        row7_pass = 0
        for col in cols:
            expected = f"={col}2+{col}5"
            cell_coord = f"{col}7"
            if check_formula(ws, cell_coord, expected):
                row7_pass += 1
                print(f"PASS: {cell_coord} has correct formula {expected}")
            else:
                actual = ws[cell_coord].value
                print(f"FAIL: {cell_coord} expected '{expected}', found '{actual}'")
        if row7_pass == 5:
            print(f"PASS: Component 3 — All Total Comp formulas correct (0.35 pts)")
            total_score += 0.35
        elif row7_pass > 0:
            partial = round(0.35 * row7_pass / 5, 4)
            print(f"PARTIAL: Component 3 — {row7_pass}/5 Total Comp formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No Total Comp formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved state, then verify
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
