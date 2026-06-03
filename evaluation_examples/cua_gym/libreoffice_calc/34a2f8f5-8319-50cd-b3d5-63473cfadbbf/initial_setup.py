"""
Initial Setup: Sales compensation plan modeler
Task ID: calc_sales_067
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CompModel ---
    ws = wb.active
    ws.title = 'CompModel'

    # Row 1: Headers
    headers = ['Component', '50%', '75%', '100%', '125%', '150%']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Row labels
    ws['A2'] = 'Base Salary'
    ws['A3'] = 'Variable (at plan)'
    ws['A4'] = 'Attainment Factor'
    ws['A5'] = 'Earned Variable'
    ws['A6'] = 'Accelerator'
    ws['A7'] = 'Total Comp'

    # Row 2: Base Salary (fixed at 80000 across all scenarios)
    for col in range(2, 7):  # B-F
        ws.cell(row=2, column=col, value=80000)

    # Row 3: Variable at plan (fixed at 60000 across all scenarios)
    for col in range(2, 7):
        ws.cell(row=3, column=col, value=60000)

    # Row 4: Attainment factors
    attainment_factors = [0.25, 0.625, 1.0, 1.5, 2.25]
    for i, factor in enumerate(attainment_factors):
        ws.cell(row=4, column=i + 2, value=factor)

    # Rows 5, 6, 7: LEFT EMPTY - these are what the agent needs to build

    # --- Formatting for readability ---
    # Header row: bold with light blue background
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Row labels: bold
    label_font = Font(bold=True, size=11)
    for row in range(2, 8):
        ws.cell(row=row, column=1).font = label_font

    # Number formatting for currency cells (rows 2, 3)
    for row in [2, 3]:
        for col in range(2, 7):
            ws.cell(row=row, column=col).number_format = '$#,##0'

    # Number formatting for attainment factor (row 4)
    for col in range(2, 7):
        ws.cell(row=4, column=col).number_format = '0.000'

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 14

    # --- Additional sheet: Assumptions ---
    ws2 = wb.create_sheet('Assumptions')
    ws2['A1'] = 'Sales Compensation Plan Assumptions'
    ws2['A1'].font = Font(bold=True, size=14)

    ws2['A3'] = 'Parameter'
    ws2['B3'] = 'Value'
    ws2['A3'].font = Font(bold=True)
    ws2['B3'].font = Font(bold=True)

    assumptions = [
        ['Annual Base Salary', 80000],
        ['On-Target Variable Pay', 60000],
        ['On-Target Earnings (OTE)', 140000],
        ['Accelerator Threshold', '100% attainment'],
        ['Below Threshold Factor', 'Linear (50% att = 0.25x, 75% att = 0.625x)'],
        ['Above Threshold Multiplier', 'Accelerated (125% att = 1.5x, 150% att = 2.25x)'],
        ['Plan Effective Date', '2025-01-01'],
        ['Review Cycle', 'Quarterly'],
    ]
    for r, (param, val) in enumerate(assumptions, 4):
        ws2.cell(row=r, column=1, value=param)
        ws2.cell(row=r, column=2, value=val)

    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 45

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
