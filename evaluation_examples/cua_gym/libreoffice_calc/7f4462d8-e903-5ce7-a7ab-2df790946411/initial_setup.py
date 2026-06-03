"""
Initial Setup: Daily Production Schedule Tracker
Task ID: calc_ops_production_schedule_026
Domain: libreoffice_calc

Creates a ProductionSchedule sheet with 90 rows of data (Date, Product Code,
Target Units pre-filled). Shift (B), Production Line (C), Actual Units (F),
Throughput % (G), and Performance Flag (H) are left empty — to be filled
by the agent as part of the task.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_production_schedule_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Seed for reproducibility
random.seed(42)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProductionSchedule ---
    ws = wb.active
    ws.title = 'ProductionSchedule'

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 18

    # Row 1: Headers
    headers = ['Date', 'Shift', 'Production Line', 'Product Code',
               'Target Units', 'Actual Units', 'Throughput %', 'Performance Flag']

    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 20

    # Product codes and target unit ranges
    products = [
        ('PRD-1001', 800, 1200),
        ('PRD-1002', 600, 900),
        ('PRD-1003', 1000, 1500),
        ('PRD-1004', 400, 700),
        ('PRD-1005', 750, 1100),
        ('PRD-1006', 500, 800),
        ('PRD-1007', 900, 1300),
        ('PRD-1008', 650, 950),
        ('PRD-1009', 350, 600),
        ('PRD-1010', 1100, 1600),
    ]

    # Generate 90 schedule entries spanning ~30 days (3 shifts per day)
    start_date = date(2025, 6, 2)

    row_num = 2
    for day_offset in range(30):
        current_date = start_date + timedelta(days=day_offset)
        for _ in range(3):  # 3 entries per day (one per shift slot)
            product = random.choice(products)
            product_code = product[0]
            target_units = random.randint(product[1], product[2])

            # Date (A), Shift (B) — empty, Production Line (C) — empty,
            # Product Code (D), Target Units (E), Actual Units (F) — empty,
            # Throughput % (G) — empty, Performance Flag (H) — empty
            ws.cell(row=row_num, column=1, value=current_date.strftime('%Y-%m-%d'))
            # B (Shift) — left empty intentionally
            # C (Production Line) — left empty intentionally
            ws.cell(row=row_num, column=4, value=product_code)
            ws.cell(row=row_num, column=5, value=target_units)
            # F, G, H — left empty intentionally

            row_num += 1

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Rows of data: 90 (rows 2 to 91)')
    print(f'  Columns with data: A (Date), D (Product Code), E (Target Units)')
    print(f'  Empty columns: B (Shift), C (Production Line), F (Actual Units), G (Throughput %), H (Performance Flag)')

create_initial()
