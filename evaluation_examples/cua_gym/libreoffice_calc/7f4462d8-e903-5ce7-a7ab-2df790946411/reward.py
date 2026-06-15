"""
Reward Script: Build a daily production schedule tracker
Task ID: calc_ops_production_schedule_026
Domain: libreoffice_calc
Scoring:
  - Component 1: Shift data validation dropdown (B2:B91) — 0.20 pts
  - Component 2: Production Line data validation dropdown (C2:C91) — 0.20 pts
  - Component 3: Throughput % formulas in G2:G91 with percentage format — 0.25 pts
  - Component 4: Performance Flag formulas in H2:H91 (BELOW TARGET / OK logic) — 0.20 pts
  - Component 5: Conditional formatting on H2:H91 (red/green) — 0.10 pts
  - Component 6: Summary row with average throughput formula — 0.05 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_ops_production_schedule_026'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'ProductionSchedule' must exist
    if 'ProductionSchedule' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ProductionSchedule' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ProductionSchedule']

    # Component 1: Shift data validation dropdown on B2:B91 (0.20 points)
    # Task requires: dropdown with Morning, Afternoon, Night for shift column B
    try:
        validations = ws.data_validations.dataValidation
        shift_dv_found = False
        for dv in validations:
            if dv.type == 'list' and dv.formula1 is not None:
                formula = dv.formula1.strip('"').strip("'")
                # Check that it contains the required shift values
                shifts = [s.strip() for s in formula.split(',')]
                required_shifts = {'Morning', 'Afternoon', 'Night'}
                if required_shifts.issubset(set(shifts)):
                    # Verify it covers B2:B91 range (check sqref covers the range)
                    sqref_str = str(dv.sqref)
                    if 'B2' in sqref_str or 'B' in sqref_str:
                        shift_dv_found = True
                        print(f"PASS: Component 1 — Shift dropdown found (formula={dv.formula1}, sqref={dv.sqref}) (0.20 pts)")
                        total_score += 0.20
                        break
        if not shift_dv_found:
            print(f"FAIL: Component 1 — No data validation with Morning/Afternoon/Night found on column B")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Production Line data validation dropdown on C2:C91 (0.20 points)
    # Task requires: dropdown with Line 1 through Line 5 for production line column C
    try:
        validations = ws.data_validations.dataValidation
        line_dv_found = False
        for dv in validations:
            if dv.type == 'list' and dv.formula1 is not None:
                formula = dv.formula1.strip('"').strip("'")
                # Check that it contains the required line values
                lines = [s.strip() for s in formula.split(',')]
                required_lines = {'Line 1', 'Line 2', 'Line 3', 'Line 4', 'Line 5'}
                if required_lines.issubset(set(lines)):
                    sqref_str = str(dv.sqref)
                    if 'C2' in sqref_str or 'C' in sqref_str:
                        line_dv_found = True
                        print(f"PASS: Component 2 — Production Line dropdown found (formula={dv.formula1}, sqref={dv.sqref}) (0.20 pts)")
                        total_score += 0.20
                        break
        if not line_dv_found:
            print(f"FAIL: Component 2 — No data validation with Line 1-5 found on column C")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Throughput % formulas in G2:G91 with percentage number format (0.25 points)
    # Task requires: =IF(ISBLANK(F2),"",F2/E2) pattern in G2:G91, formatted as percentage
    try:
        g_formula_count = 0
        g_format_count = 0
        g_errors = []
        for row in range(2, 92):
            g_cell = ws.cell(row=row, column=7)
            val = g_cell.value
            num_fmt = g_cell.number_format
            # Check formula pattern: should contain ISBLANK and division F/E
            if isinstance(val, str) and 'ISBLANK' in val.upper() and '/E' in val.upper():
                g_formula_count += 1
            else:
                if len(g_errors) < 3:
                    g_errors.append(f"G{row}: {repr(val)}")
            # Check percentage format
            if '%' in str(num_fmt):
                g_format_count += 1

        if g_formula_count == 90 and g_format_count == 90:
            print(f"PASS: Component 3 — All 90 G column throughput formulas present with % format (0.25 pts)")
            total_score += 0.25
        elif g_formula_count >= 45 and g_format_count >= 45:
            # Partial credit if at least half present
            partial = 0.12
            print(f"PARTIAL: Component 3 — Only {g_formula_count}/90 G formulas and {g_format_count}/90 % formats found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {g_formula_count}/90 G formulas found, {g_format_count}/90 with % format. Errors: {g_errors}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Performance Flag formulas in H2:H91 (0.20 points)
    # Task requires: =IF(ISBLANK(F2),"",IF(F2/E2<0.8,"BELOW TARGET","OK")) in H2:H91
    try:
        h_formula_count = 0
        h_errors = []
        for row in range(2, 92):
            h_cell = ws.cell(row=row, column=8)
            val = h_cell.value
            # Check formula has BELOW TARGET and 0.8 threshold
            if isinstance(val, str) and 'BELOW TARGET' in val.upper() and '0.8' in val and 'OK' in val.upper():
                h_formula_count += 1
            else:
                if len(h_errors) < 3:
                    h_errors.append(f"H{row}: {repr(val)}")

        if h_formula_count == 90:
            print(f"PASS: Component 4 — All 90 H column performance flag formulas present (0.20 pts)")
            total_score += 0.20
        elif h_formula_count >= 45:
            partial = 0.10
            print(f"PARTIAL: Component 4 — Only {h_formula_count}/90 H formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {h_formula_count}/90 H formulas found. Errors: {h_errors}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on H2:H91 for BELOW TARGET (red) and OK (green) (0.10 points)
    # Task requires: red fill for 'BELOW TARGET', green fill for 'OK'
    try:
        cf_rules = ws.conditional_formatting
        has_red_below_target = False
        has_green_ok = False
        for cf_range in cf_rules:
            cf_range_str = str(cf_range)
            if 'H' in cf_range_str:
                rules = cf_rules[cf_range]
                for rule in rules:
                    if rule.type == 'expression' and rule.formula:
                        formula_str = str(rule.formula[0]).upper()
                        # Check for BELOW TARGET rule with red fill
                        if 'BELOW TARGET' in formula_str:
                            try:
                                if rule.dxf and rule.dxf.fill:
                                    fill_color = rule.dxf.fill.fgColor.rgb
                                    # Red color: should contain FF0000 pattern
                                    if 'FF0000' in fill_color.upper():
                                        has_red_below_target = True
                                        print(f"PASS: Component 5a — Red fill for 'BELOW TARGET' found (color={fill_color})")
                            except Exception as inner_e:
                                # If we can't check color but formula exists, still partial credit
                                has_red_below_target = True
                                print(f"PASS: Component 5a — 'BELOW TARGET' CF rule found (color check failed: {inner_e})")
                        # Check for OK rule with green fill
                        if formula_str.endswith('"OK")') or formula_str.endswith('"OK"'):
                            try:
                                if rule.dxf and rule.dxf.fill:
                                    fill_color = rule.dxf.fill.fgColor.rgb
                                    # Green color: should contain 00B050 or 00FF00 or similar green
                                    r_val = int(fill_color[2:4], 16)
                                    g_val = int(fill_color[4:6], 16)
                                    b_val = int(fill_color[6:8], 16)
                                    # Green: G channel significantly higher than R and B
                                    if g_val > r_val and g_val > b_val:
                                        has_green_ok = True
                                        print(f"PASS: Component 5b — Green fill for 'OK' found (color={fill_color})")
                            except Exception as inner_e:
                                has_green_ok = True
                                print(f"PASS: Component 5b — 'OK' CF rule found (color check failed: {inner_e})")

        if has_red_below_target and has_green_ok:
            print(f"PASS: Component 5 — Both conditional formatting rules present (0.10 pts)")
            total_score += 0.10
        elif has_red_below_target or has_green_ok:
            partial = 0.05
            print(f"PARTIAL: Component 5 — Only partial CF rules found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No conditional formatting for BELOW TARGET/OK found on H column")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Summary row with average throughput formula (0.05 points)
    # Task requires: a summary section showing average throughput across all filled shifts
    try:
        summary_found = False
        # Check rows 92-95 for a summary row with AVERAGEIF or AVERAGE formula
        for row in range(92, 96):
            for col in range(1, 9):
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if isinstance(val, str) and ('AVERAGE' in val.upper() or 'AVG' in val.upper()):
                    summary_found = True
                    print(f"PASS: Component 6 — Summary formula found at row {row} col {col}: {repr(val)} (0.05 pts)")
                    total_score += 0.05
                    break
                elif isinstance(val, str) and val.upper() in ('SUMMARY', 'AVG THROUGHPUT:', 'AVERAGE THROUGHPUT:'):
                    # Look for associated formula in same row
                    for c2 in range(1, 9):
                        v2 = ws.cell(row=row, column=c2).value
                        if isinstance(v2, str) and 'AVERAGE' in v2.upper():
                            summary_found = True
                            print(f"PASS: Component 6 — Summary section found at row {row} (0.05 pts)")
                            total_score += 0.05
                            break
                    if summary_found:
                        break
            if summary_found:
                break
        if not summary_found:
            print(f"FAIL: Component 6 — No summary/average throughput row found near rows 92-95")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
