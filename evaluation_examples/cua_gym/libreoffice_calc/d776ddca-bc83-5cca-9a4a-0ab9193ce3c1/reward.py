"""
Reward Script: Delivery Tracking Sheet with Data Validation, IF Formulas, and Conditional Formatting
Task ID: calc_ops_supply_chain_delivery_tracking_010
Domain: libreoffice_calc
Scoring:
  - Component 1: Data validation dropdown on F2:F71 with correct options (0.35 pts)
  - Component 2: G2:G71 contain IF formula pattern for days overdue (0.35 pts)
  - Component 3: Conditional formatting on F2:F71 with correct colors (0.30 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_delivery_tracking_010'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Shipments' sheet exists
    if 'Shipments' not in wb.sheetnames:
        print("CRITICAL: 'Shipments' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Shipments']

    # Component 1: Data validation dropdown on F2:F71 (0.35 points)
    # The initial file has 0 validations; the golden file must have a list validation
    # on F2:F71 with options: In Transit, Delivered, Delayed, Returned
    try:
        validations = ws.data_validations.dataValidation
        found_dv = False
        dv_detail = "No list validation found"

        # Required options in the validation (order-insensitive, but must include all 4)
        required_options = {'In Transit', 'Delayed', 'Delivered', 'Returned'}

        for dv in validations:
            if dv.type != 'list':
                continue

            # Check formula1 contains the 4 required status options
            formula = dv.formula1 or ''
            # Remove surrounding quotes
            formula_clean = formula.strip('"')
            options_in_formula = set(opt.strip() for opt in formula_clean.split(','))

            if required_options.issubset(options_in_formula):
                # Check that the validation applies to F2:F71 range
                sqref_str = str(dv.sqref)
                # Accept any sqref that covers F2:F71 (may be written as F2:F71 or similar)
                if 'F' in sqref_str:
                    found_dv = True
                    dv_detail = f"type=list, formula1={dv.formula1}, sqref={sqref_str}"
                    break

        if found_dv:
            print(f"PASS: Component 1 — Data validation dropdown on F column with correct options ({dv_detail}) (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Expected list validation on F2:F71 with options 'In Transit,Delivered,Delayed,Returned'. {dv_detail}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G2:G71 contain IF formula for days overdue (0.35 points)
    # Formula pattern: =IF(Fn="Delayed",MAX(0,TODAY()-En),0)
    # Check that all 70 rows have a formula in column G matching this pattern
    try:
        formula_count = 0
        formula_errors = []
        expected_pattern = re.compile(
            r'=IF\(F\d+="Delayed",MAX\(0,TODAY\(\)-E\d+\),0\)',
            re.IGNORECASE
        )

        for row in range(2, 72):
            g_val = ws.cell(row=row, column=7).value
            if g_val is None:
                formula_errors.append(f"G{row}: None")
                continue
            g_str = str(g_val).strip().replace(' ', '')
            if expected_pattern.match(g_str):
                formula_count += 1
            else:
                formula_errors.append(f"G{row}: {repr(g_val)}")

        if formula_count == 70:
            print(f"PASS: Component 2 — All 70 G-column cells (G2:G71) have correct IF formula (0.35 pts)")
            total_score += 0.35
        elif formula_count >= 60:
            # Partial credit: majority of formulas correct
            print(f"PARTIAL PASS: Component 2 — {formula_count}/70 G-column cells have correct IF formula. "
                  f"Sample errors: {formula_errors[:3]}")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Only {formula_count}/70 G-column cells have correct IF formula. "
                  f"Sample errors: {formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on F2:F71 with correct colors (0.30 points)
    # Expected: red fill for 'Delayed', orange for 'In Transit', green for 'Delivered'
    # Red: FFFF0000, Orange: FFFFA500, Green: FF00B050 (or similar green)
    try:
        cf_rules = ws.conditional_formatting
        expected_rules = {
            'Delayed': 'FFFF0000',      # red
            'In Transit': 'FFFFA500',   # orange
            'Delivered': 'FF00B050',    # green
        }

        found_colors = {}  # status -> color found
        cf_on_f_column = False

        for cf in cf_rules:
            cf_str = str(cf)
            # Check if this CF applies to F column
            if 'F' not in cf_str:
                continue
            cf_on_f_column = True

            for rule in cf.rules:
                if rule.type != 'expression':
                    continue
                if not rule.formula:
                    continue
                formula_str = rule.formula[0] if rule.formula else ''

                # Identify which status this rule covers
                for status in ['Delayed', 'In Transit', 'Delivered']:
                    if status in formula_str:
                        # Check the fill color
                        if rule.dxf and rule.dxf.fill:
                            try:
                                color = rule.dxf.fill.fgColor.rgb
                                found_colors[status] = color
                            except Exception:
                                found_colors[status] = None

        if not cf_on_f_column:
            print("FAIL: Component 3 — No conditional formatting found on F column")
        else:
            # Check how many correct color mappings exist
            correct_mappings = 0
            for status, expected_color in expected_rules.items():
                found_color = found_colors.get(status)
                if found_color == expected_color:
                    correct_mappings += 1
                    print(f"  CF {status}: {found_color} matches expected {expected_color} ✓")
                else:
                    print(f"  CF {status}: found {found_color}, expected {expected_color}")

            if correct_mappings == 3:
                print(f"PASS: Component 3 — All 3 conditional formatting rules correct on F column (0.30 pts)")
                total_score += 0.30
            elif correct_mappings >= 2:
                print(f"PARTIAL PASS: Component 3 — {correct_mappings}/3 conditional formatting rules correct")
                total_score += 0.15
            elif correct_mappings >= 1:
                print(f"PARTIAL PASS: Component 3 — {correct_mappings}/3 conditional formatting rules correct")
                total_score += 0.08
            else:
                print(f"FAIL: Component 3 — No correct conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
