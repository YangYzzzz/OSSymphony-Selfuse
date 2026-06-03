"""
Initial Setup: Delivery tracking sheet with shipment data
Task ID: calc_ops_supply_chain_delivery_tracking_010
Domain: libreoffice_calc

Creates a spreadsheet with:
- Sheet 'Shipments' with 70 shipment rows
- Columns A-E filled with realistic data
- Columns F (Status) and G (Days Overdue) LEFT EMPTY
- No data validation, no conditional formatting
"""

import openpyxl
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_supply_chain_delivery_tracking_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Shipments'

    # --- Headers ---
    headers = ['Shipment ID', 'Carrier', 'Origin', 'Destination', 'Promised Date', 'Status', 'Days Overdue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Carriers ---
    carriers = [
        'FedEx', 'UPS', 'DHL', 'USPS', 'OnTrac',
        'Amazon Logistics', 'XPO Logistics', 'Estes Express', 'Old Dominion', 'SAIA'
    ]

    # --- Realistic city pairs ---
    origins = [
        'Chicago, IL', 'Los Angeles, CA', 'New York, NY', 'Houston, TX', 'Dallas, TX',
        'Atlanta, GA', 'Seattle, WA', 'Miami, FL', 'Denver, CO', 'Phoenix, AZ',
        'Boston, MA', 'Portland, OR', 'Minneapolis, MN', 'Charlotte, NC', 'Detroit, MI'
    ]
    destinations = [
        'San Francisco, CA', 'Boston, MA', 'Austin, TX', 'Nashville, TN', 'Columbus, OH',
        'Kansas City, MO', 'Salt Lake City, UT', 'Tampa, FL', 'Baltimore, MD', 'Indianapolis, IN',
        'Louisville, KY', 'Memphis, TN', 'Richmond, VA', 'Omaha, NE', 'Raleigh, NC',
        'Cleveland, OH', 'Pittsburgh, PA', 'St. Louis, MO', 'Sacramento, CA', 'New Orleans, LA'
    ]

    # --- Generate 70 shipment rows ---
    # today = 2026-03-04 (based on env date)
    base_date = date(2026, 3, 4)

    random.seed(42)  # reproducible

    for row in range(2, 72):
        idx = row - 2

        # Shipment ID
        shipment_id = f'SHP-{2026000 + idx + 1:07d}'

        # Carrier
        carrier = carriers[idx % len(carriers)]

        # Origin and destination (ensure different)
        origin = origins[idx % len(origins)]
        dest = destinations[idx % len(destinations)]

        # Promised Date — mix of past and future dates
        # Past dates (overdue) for about 40% of rows
        if idx % 5 in (0, 1):
            # Past dates: 1 to 30 days ago
            offset = -(idx % 30 + 1)
        elif idx % 5 == 2:
            # Near future: 1-7 days
            offset = idx % 7 + 1
        else:
            # Further future: 8-60 days
            offset = (idx % 53) + 8

        promised_date = base_date + timedelta(days=offset)

        # Write A-E (F and G intentionally left empty)
        ws.cell(row=row, column=1, value=shipment_id)
        ws.cell(row=row, column=2, value=carrier)
        ws.cell(row=row, column=3, value=origin)
        ws.cell(row=row, column=4, value=dest)
        ws.cell(row=row, column=5, value=promised_date)
        ws.cell(row=row, column=5).number_format = 'yyyy-mm-dd'
        # Column F (Status) — empty
        # Column G (Days Overdue) — empty

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Shipments')
    print(f'  Rows with data: 70 (rows 2-71)')
    print(f'  Columns A-E filled, F and G EMPTY (no validation, no formulas)')

create_initial()
