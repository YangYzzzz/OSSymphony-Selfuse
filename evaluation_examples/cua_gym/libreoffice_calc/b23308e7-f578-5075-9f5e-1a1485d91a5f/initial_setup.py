"""
Initial Setup: Sales analysis with pivot tables in Sheet2
Task ID: osworld_calc_pivot_multi_styled_006
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Sales Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Date', 'Customer Segment', 'Product', 'Units', 'Revenue', 'Order Value']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic sales transaction data (25 rows)
    data = [
        [date(2024, 1, 5),  'Enterprise',  'Software License',   12, 14400.00,  1200.00],
        [date(2024, 1, 12), 'SMB',         'Cloud Storage',       5,  1750.00,   350.00],
        [date(2024, 1, 18), 'Consumer',    'Mobile App',         30,  2700.00,    90.00],
        [date(2024, 1, 22), 'Enterprise',  'Cloud Storage',       8,  4800.00,   600.00],
        [date(2024, 1, 28), 'SMB',         'Software License',    3,  3600.00,  1200.00],
        [date(2024, 2, 3),  'Consumer',    'Hardware Device',    15,  7500.00,   500.00],
        [date(2024, 2, 9),  'Enterprise',  'Mobile App',         50,  4500.00,    90.00],
        [date(2024, 2, 14), 'SMB',         'Hardware Device',     7,  3500.00,   500.00],
        [date(2024, 2, 19), 'Consumer',    'Cloud Storage',      20,  2200.00,   110.00],
        [date(2024, 2, 25), 'Enterprise',  'Software License',   10, 12000.00,  1200.00],
        [date(2024, 3, 4),  'SMB',         'Mobile App',         25,  2250.00,    90.00],
        [date(2024, 3, 10), 'Consumer',    'Hardware Device',    18,  9000.00,   500.00],
        [date(2024, 3, 15), 'Enterprise',  'Cloud Storage',       6,  3600.00,   600.00],
        [date(2024, 3, 21), 'SMB',         'Software License',    4,  4800.00,  1200.00],
        [date(2024, 3, 27), 'Consumer',    'Mobile App',         35,  3150.00,    90.00],
        [date(2024, 4, 2),  'Enterprise',  'Hardware Device',    20, 10000.00,   500.00],
        [date(2024, 4, 8),  'SMB',         'Cloud Storage',       9,  3150.00,   350.00],
        [date(2024, 4, 14), 'Consumer',    'Software License',    2,  2400.00,  1200.00],
        [date(2024, 4, 20), 'Enterprise',  'Mobile App',         40,  3600.00,    90.00],
        [date(2024, 4, 26), 'SMB',         'Hardware Device',    11,  5500.00,   500.00],
        [date(2024, 5, 2),  'Consumer',    'Cloud Storage',      14,  1540.00,   110.00],
        [date(2024, 5, 8),  'Enterprise',  'Software License',    9, 10800.00,  1200.00],
        [date(2024, 5, 15), 'SMB',         'Mobile App',         22,  1980.00,    90.00],
        [date(2024, 5, 21), 'Consumer',    'Hardware Device',    16,  8000.00,   500.00],
        [date(2024, 5, 27), 'Enterprise',  'Cloud Storage',       7,  4200.00,   600.00],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Format date column
    for r in range(2, len(data) + 2):
        ws1.cell(row=r, column=1).number_format = 'yyyy-mm-dd'

    # Format revenue and order value columns
    for r in range(2, len(data) + 2):
        ws1.cell(row=r, column=5).number_format = '#,##0.00'
        ws1.cell(row=r, column=6).number_format = '#,##0.00'

    # Column widths for readability
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 14

    # --- Sheet2: Empty (agent will create pivot tables here) ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2 is intentionally empty — agent task is to build pivot tables here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
