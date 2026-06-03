"""
Reward Script: Comprehensive Sales Analysis in Sheet2 with Three Pivot Tables and Styled Header
Task ID: osworld_calc_pivot_multi_styled_006
Domain: libreoffice_calc
Scoring:
  Component 1: Merged styled header (A1:E1 merged, blue fill, bold white font) — 0.25 pts
  Component 2: Revenue by Product pivot table (3 products, correct totals) — 0.30 pts
  Component 3: Revenue by Customer Segment pivot table (correct totals) — 0.25 pts
  Component 4: Average Order Value by Month pivot table (correct averages) — 0.20 pts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_006'

# Ground truth values (derived from task requirements and golden file)
EXPECTED_PRODUCT_REVENUES = {
    'Cloud Storage': 21240,
    'Hardware Device': 43500,
    'Mobile App': 18180,
    'Software License': 48000,
}

EXPECTED_SEGMENT_REVENUES = {
    'Consumer': 36490,
    'Enterprise': 67900,
    'SMB': 26530,
}

EXPECTED_MONTHLY_AVG = {
    'January': 688,
    'February': 480,
    'March': 496,
    'April': 528,
    'May': 500,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Sheet2']

    # Precondition: Sheet2 must have meaningful content (not empty)
    if ws.max_row < 5:
        print(f"FAIL: Sheet2 is nearly empty (max_row={ws.max_row}). Task not completed.")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 1: Merged styled header at A1 (0.25 points)
    # Task requires: merged header cell with blue fill and bold white text
    # Golden: A1:E1 merged, fill fgColor = FF4472C4, font bold=True, font color = FFFFFFFF
    try:
        cell_a1 = ws['A1']

        # Check merged cell range covers A1:E1
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        header_merged = any('A1' in r for r in merged_ranges)

        # Check blue fill (blue is broadly defined — we look for a blue-ish ARGB fill)
        fill_color = None
        try:
            fill_color = cell_a1.fill.fgColor.rgb
        except Exception:
            fill_color = None

        has_blue_fill = (
            fill_color is not None
            and fill_color not in ('00000000', '00FFFFFF', 'FFFFFFFF', 'FFFFFFFF')
            and cell_a1.fill.patternType == 'solid'
        )

        # Check bold white font
        font_bold = cell_a1.font.bold is True
        font_color = None
        try:
            font_color = cell_a1.font.color.rgb
        except Exception:
            font_color = None
        has_white_font = (font_color is not None and 'FFFFFF' in font_color.upper())

        # Check header text is present (any non-empty text referencing dashboard/sales/analysis)
        header_text = cell_a1.value
        has_header_text = (
            header_text is not None
            and isinstance(header_text, str)
            and len(header_text.strip()) > 0
        )

        if header_merged and has_blue_fill and font_bold and has_white_font and has_header_text:
            print(f"PASS: Component 1 — Merged styled header present. "
                  f"Merged ranges: {merged_ranges}, fill={fill_color}, "
                  f"bold={font_bold}, font_color={font_color}, text='{header_text}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Styled header check failed. "
                  f"merged={header_merged}, has_blue_fill={has_blue_fill} (fill={fill_color}), "
                  f"bold={font_bold}, white_font={has_white_font} (font_color={font_color}), "
                  f"has_text={has_header_text}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Revenue by Product pivot table (0.30 points)
    # Task requires: total revenue aggregated by product name
    # Golden: Cloud Storage=21240, Hardware Device=43500, Mobile App=18180, Software License=48000
    try:
        # Search Sheet2 for a section labeled "Revenue by Product" and data below it
        # The section header is at row 3, data at rows 5-8 in the golden file
        # We'll scan for the label and then read the data below it
        product_section_row = None
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if isinstance(cell_val, str) and 'product' in cell_val.lower() and 'revenue' in cell_val.lower():
                product_section_row = r
                break

        if product_section_row is None:
            print("FAIL: Component 2 — 'Revenue by Product' section label not found in Sheet2")
        else:
            # Read data rows starting 2 rows below section header (skip column header row)
            found_products = {}
            for r in range(product_section_row + 2, product_section_row + 20):
                if r > ws.max_row:
                    break
                prod_cell = ws.cell(row=r, column=1).value
                rev_cell = ws.cell(row=r, column=2).value
                if prod_cell is None:
                    break
                if isinstance(prod_cell, str) and rev_cell is not None:
                    try:
                        found_products[prod_cell.strip()] = float(rev_cell)
                    except (ValueError, TypeError):
                        pass

            # Check each expected product value
            correct_products = 0
            tolerance = 1.0  # allow rounding
            for product, expected_rev in EXPECTED_PRODUCT_REVENUES.items():
                if product in found_products:
                    actual_rev = found_products[product]
                    if abs(actual_rev - expected_rev) <= tolerance:
                        correct_products += 1
                    else:
                        print(f"  FAIL: {product} revenue: expected {expected_rev}, found {actual_rev}")
                else:
                    print(f"  FAIL: {product} not found in product pivot table. Found: {list(found_products.keys())}")

            if correct_products == len(EXPECTED_PRODUCT_REVENUES):
                print(f"PASS: Component 2 — Revenue by Product pivot table correct. "
                      f"All {correct_products} products verified. (0.30 pts)")
                total_score += 0.30
            elif correct_products >= 2:
                # Partial credit not available — rubric requires full match for this component
                print(f"FAIL: Component 2 — Only {correct_products}/{len(EXPECTED_PRODUCT_REVENUES)} "
                      f"product revenues correct. Found: {found_products}")
            else:
                print(f"FAIL: Component 2 — Revenue by Product data incorrect. Found: {found_products}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Revenue by Customer Segment pivot table (0.25 points)
    # Task requires: total revenue aggregated by customer segment
    # Golden: Consumer=36490, Enterprise=67900, SMB=26530
    try:
        segment_section_row = None
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if isinstance(cell_val, str) and 'segment' in cell_val.lower() and 'revenue' in cell_val.lower():
                segment_section_row = r
                break

        if segment_section_row is None:
            print("FAIL: Component 3 — 'Revenue by Customer Segment' section label not found in Sheet2")
        else:
            # Read data rows starting 2 rows below section header (skip column header)
            found_segments = {}
            for r in range(segment_section_row + 2, segment_section_row + 20):
                if r > ws.max_row:
                    break
                seg_cell = ws.cell(row=r, column=1).value
                rev_cell = ws.cell(row=r, column=2).value
                if seg_cell is None:
                    break
                if isinstance(seg_cell, str) and rev_cell is not None:
                    try:
                        found_segments[seg_cell.strip()] = float(rev_cell)
                    except (ValueError, TypeError):
                        pass

            correct_segments = 0
            tolerance = 1.0
            for segment, expected_rev in EXPECTED_SEGMENT_REVENUES.items():
                if segment in found_segments:
                    actual_rev = found_segments[segment]
                    if abs(actual_rev - expected_rev) <= tolerance:
                        correct_segments += 1
                    else:
                        print(f"  FAIL: {segment} revenue: expected {expected_rev}, found {actual_rev}")
                else:
                    print(f"  FAIL: {segment} not found in segment pivot table. Found: {list(found_segments.keys())}")

            if correct_segments == len(EXPECTED_SEGMENT_REVENUES):
                print(f"PASS: Component 3 — Revenue by Customer Segment pivot table correct. "
                      f"All {correct_segments} segments verified. (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Only {correct_segments}/{len(EXPECTED_SEGMENT_REVENUES)} "
                      f"segment revenues correct. Found: {found_segments}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Average Order Value by Month pivot table (0.20 points)
    # Task requires: average order value aggregated by month
    # Golden: January=688, February=480, March=496, April=528, May=500
    try:
        month_section_row = None
        for r in range(1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if isinstance(cell_val, str) and 'month' in cell_val.lower() and (
                    'order' in cell_val.lower() or 'average' in cell_val.lower() or 'avg' in cell_val.lower()):
                month_section_row = r
                break

        if month_section_row is None:
            print("FAIL: Component 4 — 'Average Order Value by Month' section label not found in Sheet2")
        else:
            # Read data rows starting 2 rows below section header (skip column header)
            found_months = {}
            for r in range(month_section_row + 2, month_section_row + 20):
                if r > ws.max_row:
                    break
                month_cell = ws.cell(row=r, column=1).value
                avg_cell = ws.cell(row=r, column=2).value
                if month_cell is None:
                    break
                if isinstance(month_cell, str) and avg_cell is not None:
                    try:
                        found_months[month_cell.strip()] = float(avg_cell)
                    except (ValueError, TypeError):
                        pass

            correct_months = 0
            tolerance = 1.0  # allow rounding (avg may be float)
            for month, expected_avg in EXPECTED_MONTHLY_AVG.items():
                if month in found_months:
                    actual_avg = found_months[month]
                    if abs(actual_avg - expected_avg) <= tolerance:
                        correct_months += 1
                    else:
                        print(f"  FAIL: {month} avg order value: expected {expected_avg}, found {actual_avg}")
                else:
                    print(f"  FAIL: {month} not found in month pivot table. Found: {list(found_months.keys())}")

            if correct_months == len(EXPECTED_MONTHLY_AVG):
                print(f"PASS: Component 4 — Average Order Value by Month pivot table correct. "
                      f"All {correct_months} months verified. (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Only {correct_months}/{len(EXPECTED_MONTHLY_AVG)} "
                      f"monthly averages correct. Found: {found_months}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
