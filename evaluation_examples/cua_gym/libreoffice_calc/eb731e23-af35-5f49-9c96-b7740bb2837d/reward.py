"""
Reward Script: Flag employees outside salary band using VLOOKUP + IF formulas
Task ID: calc_fin_salary_band_025
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: New headers (D1='Min Band', E1='Max Band', F1='Status')        — 0.15 pts
  Component 2: VLOOKUP formulas in D2:D50 (Min Band) and E2:E50 (Max Band)   — 0.25 pts
  Component 3: IF status formulas in F2:F50 (Below/In/Above Range)            — 0.25 pts
  Component 4: Summary row at row 52 with COUNTIF for each status category    — 0.15 pts
  Component 5: Conditional formatting (Below Range=blue, Above Range=red)     — 0.20 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fin_salary_band_025'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — gate on file accessibility
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Employees' sheet must exist
    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: 'Employees' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Employees']

    # -------------------------------------------------------------------------
    # Component 1: New column headers (D1='Min Band', E1='Max Band', F1='Status')
    # (0.15 points)
    # Initial file: D1/E1/F1 are all None (no extra columns)
    # Golden file: D1='Min Band', E1='Max Band', F1='Status'
    # -------------------------------------------------------------------------
    try:
        d1 = ws.cell(row=1, column=4).value
        e1 = ws.cell(row=1, column=5).value
        f1 = ws.cell(row=1, column=6).value

        headers_ok = (
            d1 is not None and str(d1).strip() == 'Min Band' and
            e1 is not None and str(e1).strip() == 'Max Band' and
            f1 is not None and str(f1).strip() == 'Status'
        )

        if headers_ok:
            print(f"PASS: Component 1 — Headers D1='Min Band', E1='Max Band', F1='Status' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Expected headers Min Band/Max Band/Status, found: D1={repr(d1)}, E1={repr(e1)}, F1={repr(f1)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: VLOOKUP formulas in D2:D50 (Min Band) and E2:E50 (Max Band)
    # (0.25 points)
    # Initial file: columns D and E don't exist in rows 2-50
    # Golden file: VLOOKUP formulas referencing PayScale lookup table
    # Award 0.15 for D column VLOOKUPs, 0.10 for E column VLOOKUPs
    # -------------------------------------------------------------------------
    try:
        d_vlookup_count = 0
        e_vlookup_count = 0
        total_rows = 49  # rows 2 to 50

        for row in range(2, 51):
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value

            if d_val and isinstance(d_val, str) and 'VLOOKUP' in d_val.upper() and 'PAYSCALE' in d_val.upper():
                d_vlookup_count += 1
            if e_val and isinstance(e_val, str) and 'VLOOKUP' in e_val.upper() and 'PAYSCALE' in e_val.upper():
                e_vlookup_count += 1

        # Full credit if all 49 rows have VLOOKUP formulas in D and E
        d_ok = d_vlookup_count >= total_rows
        e_ok = e_vlookup_count >= total_rows

        if d_ok and e_ok:
            print(f"PASS: Component 2 — VLOOKUP in D2:D50 ({d_vlookup_count}/49) and E2:E50 ({e_vlookup_count}/49) (0.25 pts)")
            total_score += 0.25
        elif d_ok:
            print(f"PASS (partial): Component 2a — VLOOKUP in D2:D50 ({d_vlookup_count}/49) (0.15 pts)")
            total_score += 0.15
            print(f"FAIL: Component 2b — VLOOKUP in E2:E50 only {e_vlookup_count}/49 found")
        elif e_ok:
            print(f"FAIL: Component 2a — VLOOKUP in D2:D50 only {d_vlookup_count}/49 found")
            print(f"PASS (partial): Component 2b — VLOOKUP in E2:E50 ({e_vlookup_count}/49) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — VLOOKUP formulas missing/incomplete: D={d_vlookup_count}/49, E={e_vlookup_count}/49")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: IF status formulas in F2:F50
    # (0.25 points)
    # Initial file: column F doesn't exist in rows 2-50
    # Golden file: =IF(C<D,"Below Range",IF(C>E,"Above Range","In Range"))
    # -------------------------------------------------------------------------
    try:
        f_if_count = 0
        f_below_count = 0
        f_above_count = 0

        for row in range(2, 51):
            f_val = ws.cell(row=row, column=6).value
            if f_val and isinstance(f_val, str):
                f_upper = f_val.upper()
                if 'IF(' in f_upper:
                    f_if_count += 1
                    if 'BELOW RANGE' in f_upper:
                        f_below_count += 1
                    if 'ABOVE RANGE' in f_upper:
                        f_above_count += 1

        # Full credit: all 49 rows have IF formulas with "Below Range" and "Above Range"
        if f_if_count >= 49 and f_below_count >= 49 and f_above_count >= 49:
            print(f"PASS: Component 3 — IF status formulas in F2:F50 with Below/Above/In Range (0.25 pts)")
            total_score += 0.25
        elif f_if_count >= 49:
            print(f"PASS (partial): Component 3 — IF formulas present ({f_if_count}/49) but missing status text (Below={f_below_count}, Above={f_above_count}) (0.15 pts)")
            total_score += 0.15
        elif f_if_count > 0:
            print(f"FAIL (partial): Component 3 — IF formulas only {f_if_count}/49 rows (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No IF formulas found in F2:F50")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Summary row at row 52 with COUNTIF for each status category
    # (0.15 points)
    # Initial file: rows only go up to 50 (49 data rows)
    # Golden file: row 52 has Summary label and COUNTIF formulas
    # -------------------------------------------------------------------------
    try:
        a52 = ws.cell(row=52, column=1).value
        c52 = ws.cell(row=52, column=3).value
        e52 = ws.cell(row=52, column=5).value
        g52 = ws.cell(row=52, column=7).value

        summary_label_ok = a52 is not None and str(a52).strip().lower() == 'summary'

        c52_ok = bool(c52 and isinstance(c52, str) and
                      'COUNTIF' in c52.upper() and 'BELOW RANGE' in c52.upper())
        e52_ok = bool(e52 and isinstance(e52, str) and
                      'COUNTIF' in e52.upper() and 'IN RANGE' in e52.upper())
        g52_ok = bool(g52 and isinstance(g52, str) and
                      'COUNTIF' in g52.upper() and 'ABOVE RANGE' in g52.upper())

        countif_count = sum([c52_ok, e52_ok, g52_ok])

        if summary_label_ok and countif_count == 3:
            print(f"PASS: Component 4 — Summary row 52 with COUNTIF for all 3 categories (0.15 pts)")
            total_score += 0.15
        elif summary_label_ok and countif_count > 0:
            partial = round(0.05 * countif_count + 0.02, 2)
            print(f"PASS (partial): Component 4 — Summary label present, {countif_count}/3 COUNTIF formulas ({partial} pts)")
            total_score += partial
        elif countif_count == 3:
            print(f"PASS (partial): Component 4 — All 3 COUNTIF formulas present but no Summary label (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Summary row missing/incomplete: label={repr(a52)}, COUNTIF count={countif_count}/3")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Conditional formatting rules (Below Range=blue, Above Range=red)
    # (0.20 points)
    # Initial file: no conditional formatting on Employees sheet
    # Golden file: CF rules on F2:F50 for Below Range (blue ~FF63B3D1) and Above Range (red ~FFFF6B6B)
    # -------------------------------------------------------------------------
    try:
        def _is_blue_fill(dxf_fill):
            """Return True if dxf fill color is blue-ish (B or G > R), else False."""
            try:
                fg_color = dxf_fill.fgColor.rgb
                if fg_color and len(fg_color) == 8:
                    r = int(fg_color[2:4], 16)
                    g = int(fg_color[4:6], 16)
                    b = int(fg_color[6:8], 16)
                    return b > r or g > r
                return False
            except Exception:
                return False

        def _is_red_fill(dxf_fill):
            """Return True if dxf fill color is red-ish (R > G and R > B), else False."""
            try:
                fg_color = dxf_fill.fgColor.rgb
                if fg_color and len(fg_color) == 8:
                    r = int(fg_color[2:4], 16)
                    g = int(fg_color[4:6], 16)
                    b = int(fg_color[6:8], 16)
                    return r > g and r > b
                return False
            except Exception:
                return False

        cf_rules = ws.conditional_formatting
        cf_list = list(cf_rules)

        below_found_count = 0
        above_found_count = 0

        for cf_range in cf_list:
            for rule in cf_rules[cf_range]:
                formula_str = ''
                if hasattr(rule, 'formula') and rule.formula:
                    formula_str = ' '.join(str(f) for f in rule.formula).upper()

                # Check for Below Range conditional formatting with blue-ish fill
                if ('BELOW RANGE' in formula_str and rule.dxf and rule.dxf.fill
                        and _is_blue_fill(rule.dxf.fill)):
                    try:
                        fg_color = rule.dxf.fill.fgColor.rgb
                        print(f"  Found Below Range CF with color #{fg_color}")
                    except Exception:
                        print(f"  Found Below Range CF (color read error)")
                    below_found_count += 1

                # Check for Above Range conditional formatting with red-ish fill
                if ('ABOVE RANGE' in formula_str and rule.dxf and rule.dxf.fill
                        and _is_red_fill(rule.dxf.fill)):
                    try:
                        fg_color = rule.dxf.fill.fgColor.rgb
                        print(f"  Found Above Range CF with color #{fg_color}")
                    except Exception:
                        print(f"  Found Above Range CF (color read error)")
                    above_found_count += 1

        below_range_cf = below_found_count > 0
        above_range_cf = above_found_count > 0

        if below_range_cf and above_range_cf:
            print(f"PASS: Component 5 — Both CF rules present (Below Range=blue, Above Range=red) (0.20 pts)")
            total_score += 0.20
        elif below_range_cf:
            print(f"PASS (partial): Component 5 — Below Range CF rule found, Above Range missing (0.10 pts)")
            total_score += 0.10
        elif above_range_cf:
            print(f"PASS (partial): Component 5 — Above Range CF rule found, Below Range missing (0.10 pts)")
            total_score += 0.10
        else:
            cf_count = len(cf_list)
            print(f"FAIL: Component 5 — No matching CF rules found (total CF ranges: {cf_count})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
