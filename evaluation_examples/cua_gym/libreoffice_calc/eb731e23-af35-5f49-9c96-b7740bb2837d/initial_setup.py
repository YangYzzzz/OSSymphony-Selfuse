"""
Initial Setup: Salary band compliance tracking spreadsheet
Task ID: calc_fin_salary_band_025
Domain: libreoffice_calc

Creates a workbook with:
- 'Employees' sheet: 49 employees with Name, Job Grade, Current Salary (NO band columns yet)
- 'PayScale' sheet: grades G1-G7 with min/max salary ranges
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_salary_band_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Employees ----
    ws_emp = wb.active
    ws_emp.title = 'Employees'

    # Headers (row 1) — NOT bold in initial (bold is part of task)
    headers = ['Name', 'Job Grade', 'Current Salary']
    for col, h in enumerate(headers, 1):
        ws_emp.cell(row=1, column=col, value=h)

    # 49 employee rows (rows 2-50), mix of grades G1-G7
    # Pay-scale bands (for generating realistic salaries — some in band, some outliers):
    # G1: 28000-38000, G2: 35000-48000, G3: 44000-58000, G4: 55000-72000,
    # G5: 68000-88000, G6: 84000-108000, G7: 105000-135000

    employees = [
        ('Amelia Torres',      'G3',  52400),
        ('Benjamin Okafor',    'G1',  31200),
        ('Catherine Liu',      'G5',  76800),
        ('David Schneider',    'G4',  69500),
        ('Elena Vasquez',      'G2',  42300),
        ('Frank Nakamura',     'G6',  92000),
        ('Grace Mensah',       'G3',  44600),
        ('Henry Blackwood',    'G7', 128000),
        ('Isabella Costa',     'G2',  37900),
        ('James O\'Brien',     'G4',  57300),
        ('Karen Johansson',    'G5',  90500),   # above G5 max 88000
        ('Liam Petrov',        'G1',  25800),   # below G1 min 28000
        ('Maria Fernandez',    'G6', 105200),
        ('Nathan Osei',        'G3',  58100),   # above G3 max 58000
        ('Olivia Hartmann',    'G4',  63200),
        ('Patrick Dumont',     'G2',  48800),   # above G2 max 48000
        ('Quinn Nakagawa',     'G5',  70000),
        ('Rachel Kim',         'G7', 110000),
        ('Samuel Nwosu',       'G1',  33500),
        ('Tanya Bergstrom',    'G6',  86000),
        ('Umar Hassan',        'G3',  47900),
        ('Valentina Greco',    'G4',  60100),
        ('William Adeyemi',    'G2',  36400),
        ('Xin Zhao',           'G5',  78900),
        ('Yasmin Al-Rashid',   'G6', 110000),
        ('Zachary Thorn',      'G7', 102000),   # below G7 min 105000
        ('Aisha Patel',        'G1',  29800),
        ('Bruno Martinez',     'G4',  72800),   # above G4 max 72000
        ('Clara Hoffmann',     'G3',  53700),
        ('Daniel Sousa',       'G6',  95000),
        ('Emma Carlsson',      'G2',  40100),
        ('Felix Wagner',       'G5',  68500),
        ('Gina Esposito',      'G7', 120000),
        ('Hugo Leroy',         'G1',  38500),   # above G1 max 38000
        ('Ifeoma Chukwu',      'G3',  45000),
        ('Julian Reyes',       'G4',  58800),
        ('Katrina Volkov',     'G2',  43600),
        ('Lorenzo Bianchi',    'G5',  82000),
        ('Mia Andersen',       'G6',  84500),
        ('Noel Dufour',        'G7', 130000),
        ('Oluwaseun Adebayo',  'G1',  30500),
        ('Priya Sharma',       'G3',  56000),
        ('Quentin Lambert',    'G4',  67400),
        ('Rosa Delgado',       'G5',  73200),
        ('Sven Nilsson',       'G6', 100000),
        ('Tomoko Inoue',       'G7', 115000),
        ('Uma Krishnan',       'G2',  39800),
        ('Viktor Kowalski',    'G4',  55500),
        ('Wendy Okonkwo',      'G3',  49200),
    ]

    for r, (name, grade, salary) in enumerate(employees, 2):
        ws_emp.cell(row=r, column=1, value=name)
        ws_emp.cell(row=r, column=2, value=grade)
        ws_emp.cell(row=r, column=3, value=salary)

    # ---- Sheet 2: PayScale ----
    ws_pay = wb.create_sheet('PayScale')

    # Headers
    pay_headers = ['Grade', 'Min Salary', 'Max Salary']
    for col, h in enumerate(pay_headers, 1):
        ws_pay.cell(row=1, column=col, value=h)

    # Grade bands (G1-G7)
    pay_data = [
        ('G1',  28000,  38000),
        ('G2',  35000,  48000),
        ('G3',  44000,  58000),
        ('G4',  55000,  72000),
        ('G5',  68000,  88000),
        ('G6',  84000, 108000),
        ('G7', 105000, 135000),
    ]

    for r, (grade, min_sal, max_sal) in enumerate(pay_data, 2):
        ws_pay.cell(row=r, column=1, value=grade)
        ws_pay.cell(row=r, column=2, value=min_sal)
        ws_pay.cell(row=r, column=3, value=max_sal)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Employees sheet: 49 employees (rows 2-50), 3 columns')
    print(f'  PayScale sheet: grades G1-G7 (rows 2-8), 3 columns')


create_initial()
