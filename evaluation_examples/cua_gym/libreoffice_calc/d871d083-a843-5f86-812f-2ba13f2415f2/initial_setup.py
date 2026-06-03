"""
Initial Setup: Create web_analytics.xlsx with 365 rows of daily channel traffic data
Task ID: calc_gg5_044
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traffic"

    # --- Headers ---
    headers = ["Date", "Page", "Channel", "Sessions", "Avg Session Duration", "Bounce Rate", "Conversions"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 34

    # --- Data generation ---
    channels = ["Organic Search", "Paid Search", "Social Media", "Direct", "Referral", "Email"]
    pages = [
        "/home", "/products", "/about", "/contact", "/blog/seo-tips",
        "/blog/marketing-guide", "/pricing", "/solutions/enterprise",
        "/solutions/small-business", "/case-studies", "/resources/whitepaper",
        "/blog/analytics-101", "/demo", "/support", "/careers",
        "/partners", "/blog/content-strategy", "/features", "/integrations",
        "/blog/social-media-trends"
    ]

    start_date = datetime(2025, 1, 1)

    for row_idx in range(2, 367):  # rows 2-366 = 365 data rows
        day_offset = row_idx - 2
        current_date = start_date + timedelta(days=day_offset)
        channel = random.choice(channels)
        page = random.choice(pages)

        # Sessions: varies by channel
        if channel == "Organic Search":
            sessions = random.randint(120, 850)
        elif channel == "Paid Search":
            sessions = random.randint(80, 500)
        elif channel == "Social Media":
            sessions = random.randint(50, 400)
        elif channel == "Direct":
            sessions = random.randint(100, 600)
        elif channel == "Referral":
            sessions = random.randint(30, 250)
        else:  # Email
            sessions = random.randint(40, 300)

        # Avg Session Duration (seconds): 30-600
        avg_duration = round(random.uniform(30.0, 600.0), 1)

        # Bounce Rate: 0.10 - 0.85
        bounce_rate = round(random.uniform(0.10, 0.85), 4)

        # Conversions: proportional to sessions
        conversions = random.randint(0, max(1, sessions // 10))

        ws.cell(row=row_idx, column=1, value=current_date)
        ws.cell(row=row_idx, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=row_idx, column=2, value=page)
        ws.cell(row=row_idx, column=3, value=channel)
        ws.cell(row=row_idx, column=4, value=sessions)
        ws.cell(row=row_idx, column=5, value=avg_duration)
        ws.cell(row=row_idx, column=6, value=bounce_rate)
        ws.cell(row=row_idx, column=6).number_format = '0.00%'
        ws.cell(row=row_idx, column=7, value=conversions)

    # H1 and H2 must be EMPTY (task requires placing content there)
    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter on data range
    ws.auto_filter.ref = "A1:G366"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
