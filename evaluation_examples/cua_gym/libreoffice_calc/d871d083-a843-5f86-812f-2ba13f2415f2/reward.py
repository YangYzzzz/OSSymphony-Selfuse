"""
Reward Script: AVERAGEIFS formula for average session duration
Task ID: calc_gg5_044
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.3): H1 has a descriptive label
  - Component 2 (0.4): H2 contains AVERAGEIFS formula with correct structure
  - Component 3 (0.3): H2 formula references correct ranges and both criteria
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_044'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Traffic sheet must exist
    if 'Traffic' not in wb.sheetnames:
        print("CRITICAL: 'Traffic' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Traffic']

    # Component 1: H1 contains a descriptive label (0.3 points)
    # This should be empty in initial_env and have a label in golden_env
    try:
        h1_val = ws['H1'].value
        if h1_val is not None and isinstance(h1_val, str) and len(h1_val.strip()) >= 5:
            # Check that the label is relevant (mentions duration, organic, or bounce)
            h1_lower = h1_val.lower()
            relevant_keywords = ['duration', 'organic', 'bounce', 'avg', 'average', 'session']
            keyword_matches = sum(1 for kw in relevant_keywords if kw in h1_lower)
            if keyword_matches >= 2:
                print(f"PASS: Component 1 — H1 has relevant label: '{h1_val}' (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 — H1 label '{h1_val}' lacks relevant keywords (need >=2 of {relevant_keywords})")
        else:
            print(f"FAIL: Component 1 — H1 is empty or too short: {repr(h1_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: H2 contains an AVERAGEIFS formula (0.4 points)
    # Must be a formula string starting with =AVERAGEIFS
    try:
        h2_val = ws['H2'].value
        if h2_val is not None and isinstance(h2_val, str):
            h2_clean = h2_val.strip().upper().replace(" ", "")
            if h2_clean.startswith("=AVERAGEIFS("):
                print(f"PASS: Component 2 — H2 contains AVERAGEIFS formula: '{h2_val}' (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — H2 has formula but not AVERAGEIFS: '{h2_val}'")
        else:
            print(f"FAIL: Component 2 — H2 is empty or not a formula: {repr(h2_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: H2 formula references correct criteria — both "Organic Search" and "<0.5" (0.3 points)
    # The formula must reference column E (avg_duration), column C (channel), and column F (bounce rate)
    try:
        h2_val = ws['H2'].value
        if h2_val is not None and isinstance(h2_val, str) and h2_val.strip().upper().startswith("=AVERAGEIFS"):
            h2_upper = h2_val.upper().replace(" ", "")

            checks_passed = 0

            # Check 1: References E column for avg_duration (the value range)
            if re.search(r'E\d+:E\d+', h2_upper):
                checks_passed += 1
            else:
                print(f"  DETAIL: H2 formula missing E column range reference")

            # Check 2: References C column for Channel criteria
            if re.search(r'C\d+:C\d+', h2_upper):
                checks_passed += 1
            else:
                print(f"  DETAIL: H2 formula missing C column range reference")

            # Check 3: References "Organic Search" criterion
            if 'ORGANICSEARCH' in h2_upper.replace('"', '').replace("'", ""):
                checks_passed += 1
            else:
                print(f"  DETAIL: H2 formula missing 'Organic Search' criterion")

            # Check 4: References F column for Bounce Rate criteria
            if re.search(r'F\d+:F\d+', h2_upper):
                checks_passed += 1
            else:
                print(f"  DETAIL: H2 formula missing F column range reference")

            # Check 5: References "<0.5" criterion for bounce rate
            if '<0.5' in h2_val.replace(" ", "") or '"<0.5"' in h2_val.replace(" ", ""):
                checks_passed += 1
            else:
                print(f"  DETAIL: H2 formula missing '<0.5' bounce rate criterion")

            if checks_passed >= 5:
                print(f"PASS: Component 3 — H2 formula has all correct references and criteria ({checks_passed}/5 sub-checks) (0.3 pts)")
                total_score += 0.3
            elif checks_passed >= 3:
                partial = round(0.3 * (checks_passed / 5), 2)
                print(f"PARTIAL: Component 3 — H2 formula has {checks_passed}/5 correct references ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — H2 formula only has {checks_passed}/5 correct references")
        else:
            print(f"FAIL: Component 3 — H2 does not contain an AVERAGEIFS formula, skipping criteria check")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
