"""
Initial Setup: Pie chart market share spreadsheet (no custom slice colors)
Task ID: calc_chart_pie_slice_color_036
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.chart import PieChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_pie_slice_color_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: MarketShare ---
    ws = wb.active
    ws.title = 'MarketShare'

    # Headers
    ws['A1'] = 'Segment'
    ws['B1'] = 'Share %'

    # Data rows (realistic market share content)
    ws['A2'] = 'Our Brand'
    ws['B2'] = 38
    ws['A3'] = 'Competitors'
    ws['B3'] = 45
    ws['A4'] = 'Unbranded'
    ws['B4'] = 17

    # Column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 12

    # --- Pie Chart with default colors (no custom DataPoint colors) ---
    pie = PieChart()
    pie.title = 'Market Share Distribution'

    # Data range includes header row for legend labels
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=4)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=4)
    pie.add_data(data_ref, titles_from_data=True)
    pie.set_categories(cats_ref)

    # Position the chart on the sheet
    ws.add_chart(pie, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Contents:')
    print('  Sheet: MarketShare')
    print('  Data: Segment/Share% with Our Brand(38), Competitors(45), Unbranded(17)')
    print('  Pie chart with default colors (no custom slice colors)')


create_initial()
