"""
Reward Script: Add comments to 8 critical formula cells and flag hardcoded values
Task ID: calc_gen_comments_057
Domain: libreoffice_calc
Scoring:
  Component 1: Comments added to the 8 formula cells (C5, D5, E5, C25, D25, C30, D30, E30) — 0.6 pts
  Component 2: Comment added to B2 flagging the hardcoded discount rate — 0.2 pts
  Component 3: C30 comment flags the hardcoded 0.10 rate with recommendation — 0.2 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_comments_057'

# The 8 critical formula cells that must have comments
FORMULA_CELLS = ['C5', 'D5', 'E5', 'C25', 'D25', 'C30', 'D30', 'E30']

# The hardcoded-value cell that must have a flagging comment
HARDCODED_VALUE_CELL = 'B2'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Checks that comments were added to 8 formula cells and hardcoded value cells.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that the Budget sheet exists
    if 'Budget' not in wb.sheetnames:
        print("CRITICAL: 'Budget' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget']

    # Component 1: Comments on 8 formula cells (0.6 points)
    # Each formula cell comment is worth 0.075 points (8 cells x 0.075 = 0.6)
    # A comment must be non-empty and contain some explanation text.
    # This FAILS on initial (0 comments) and PASSES on golden (8 comments)
    try:
        commented_formula_cells = []
        missing_formula_comments = []

        for coord in FORMULA_CELLS:
            cell = ws[coord]
            if cell.comment and cell.comment.text and len(cell.comment.text.strip()) > 10:
                commented_formula_cells.append(coord)
            else:
                missing_formula_comments.append(coord)

        cells_with_comments = len(commented_formula_cells)
        points_per_cell = 0.6 / len(FORMULA_CELLS)  # 0.075 per cell
        component1_score = round(cells_with_comments * points_per_cell, 4)

        if cells_with_comments == len(FORMULA_CELLS):
            print(f"PASS: Component 1 — All {len(FORMULA_CELLS)} formula cells have comments ({component1_score:.3f} pts)")
            print(f"  Commented cells: {commented_formula_cells}")
            total_score += component1_score
        elif cells_with_comments > 0:
            print(f"PARTIAL: Component 1 — {cells_with_comments}/{len(FORMULA_CELLS)} formula cells have comments ({component1_score:.3f} pts)")
            print(f"  Commented: {commented_formula_cells}")
            print(f"  Missing: {missing_formula_comments}")
            total_score += component1_score
        else:
            print(f"FAIL: Component 1 — No formula cells have comments (0.0 pts)")
            print(f"  Missing comments on all: {FORMULA_CELLS}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check formula cell comments: {e}")

    # Component 2: Comment on B2 flagging the hardcoded discount rate (0.2 points)
    # B2 contains 0.15 (hardcoded discount rate). A comment should flag this as a
    # hardcoded parameter and suggest moving it to an assumptions section.
    # This FAILS on initial (no comment) and PASSES on golden (comment added)
    try:
        cell_b2 = ws[HARDCODED_VALUE_CELL]
        b2_has_comment = (cell_b2.comment is not None and
                          cell_b2.comment.text and
                          len(cell_b2.comment.text.strip()) > 10)

        if b2_has_comment:
            comment_text_lower = cell_b2.comment.text.lower()
            # Check that the comment mentions the parameter/assumption aspect
            has_parameter_content = any(keyword in comment_text_lower for keyword in
                                        ['parameter', 'assumption', 'hardcoded', 'hard-coded',
                                         'discount', 'rate', 'assumption', 'move'])
            if has_parameter_content:
                print(f"PASS: Component 2 — B2 has a comment flagging the hardcoded discount rate (0.2 pts)")
                print(f"  Comment preview: {cell_b2.comment.text[:100].strip()!r}")
                total_score += 0.2
            else:
                print(f"PARTIAL: Component 2 — B2 has a comment but it doesn't clearly flag the hardcoded nature (0.1 pts)")
                print(f"  Comment preview: {cell_b2.comment.text[:100].strip()!r}")
                total_score += 0.1
        else:
            print(f"FAIL: Component 2 — B2 has no comment (0.0 pts)")
            print(f"  B2 should have a comment flagging the hardcoded 0.15 discount rate")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check B2 comment: {e}")

    # Component 3: C30 comment flags the hardcoded 0.10 discount rate (0.2 points)
    # C30 contains =NPV(0.10,C25:C29) with a hardcoded 0.10 rate.
    # The comment should note this and suggest moving it to an assumptions section.
    # This FAILS on initial (no comment) and PASSES on golden (comment with hardcoded flag)
    try:
        cell_c30 = ws['C30']
        c30_has_comment = (cell_c30.comment is not None and
                           cell_c30.comment.text and
                           len(cell_c30.comment.text.strip()) > 10)

        if c30_has_comment:
            comment_text_lower = cell_c30.comment.text.lower()
            # Check that the comment mentions the hardcoded 0.10 rate
            has_hardcoded_flag = any(keyword in comment_text_lower for keyword in
                                     ['hardcoded', 'hard-coded', 'hard coded', '0.10', '10%',
                                      'parameter', 'assumption', 'recommendation'])
            if has_hardcoded_flag:
                print(f"PASS: Component 3 — C30 comment flags the hardcoded 0.10 NPV discount rate (0.2 pts)")
                print(f"  Comment preview: {cell_c30.comment.text[:120].strip()!r}")
                total_score += 0.2
            else:
                print(f"PARTIAL: Component 3 — C30 has a comment but doesn't flag the hardcoded rate (0.1 pts)")
                print(f"  Comment preview: {cell_c30.comment.text[:120].strip()!r}")
                total_score += 0.1
        else:
            print(f"FAIL: Component 3 — C30 has no comment (0.0 pts)")
            print(f"  C30 should have a comment flagging the hardcoded 0.10 discount rate in NPV formula")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check C30 comment: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
