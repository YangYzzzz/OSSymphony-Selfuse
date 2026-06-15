"""
Initial Setup: Budget spreadsheet with formula cells needing comments
Task ID: calc_gen_comments_057
Domain: libreoffice_calc

Creates a Budget workbook with two sheets:
- 'Budget': Contains financial data with several complex formulas (no comments yet)
- 'Rates': Currency conversion rate table used by VLOOKUP in the Budget sheet

The agent's task is to add explanatory comments to the 8 most critical formula
cells and flag hardcoded values for parameterization.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_comments_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ─── Sheet 1: Budget ───────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Budget'

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Row 1: Title
    ws['A1'] = 'Annual Budget Model — FY 2025'
    ws['A1'].font = Font(bold=True, size=14)

    # Row 2: Discount rate (hardcoded — B2)
    ws['A2'] = 'Discount Rate'
    ws['B2'] = 0.15
    ws['B2'].number_format = '0.00%'

    # Row 3: blank separator

    # Row 4: Headers
    headers = ['Product Line', 'Unit Price', 'Revenue (Base)', 'Revenue (Net)', 'Revenue (USD)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # Row 5: Summary row (contains the critical formula cells C5, D5, E5)
    ws['A5'] = 'USD'          # used as lookup key in VLOOKUP(A5, Rates!A:B, 2, 0)
    ws['B5'] = None            # not used in formulas
    ws['C5'] = '=SUM(C8:C22)*1.08'
    ws['D5'] = '=C5*(1-$B$2)'
    ws['E5'] = "=D5/VLOOKUP(A5,'Rates'!A:B,2,FALSE)"

    ws['C5'].number_format = '#,##0.00'
    ws['D5'].number_format = '#,##0.00'
    ws['E5'].number_format = '#,##0.00'

    ws['A5'].font = Font(bold=True)
    ws['C5'].font = Font(bold=True)
    ws['D5'].font = Font(bold=True)
    ws['E5'].font = Font(bold=True)

    # Row 6: blank separator

    # Row 7: Sub-headers for data rows
    sub_headers = ['Product Line', 'Unit Price', 'Units Sold', 'Weight Factor', '']
    for col, h in enumerate(sub_headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = Font(bold=True, italic=True)

    # Rows 8–22: Product data (C8:C22 = Revenue, D8:D22 = Weight Factors)
    products = [
        ('Analytics Suite',     299.00, 312,   0.12),
        ('CRM Professional',    449.00, 198,   0.15),
        ('Data Warehouse',      899.00,  87,   0.18),
        ('E-Commerce Engine',   349.00, 264,   0.11),
        ('Finance Module',      599.00, 143,   0.14),
        ('HRMS Platform',       499.00, 176,   0.13),
        ('Integration Hub',     799.00,  95,   0.16),
        ('Logistics Tracker',   279.00, 331,   0.10),
        ('Mobile SDK',          199.00, 487,   0.08),
        ('Network Monitor',     649.00, 118,   0.14),
        ('Operations Console',  549.00, 152,   0.13),
        ('Pipeline Manager',    729.00, 104,   0.16),
        ('Query Builder',       249.00, 389,   0.09),
        ('Reporting Studio',    379.00, 231,   0.11),
        ('Security Gateway',    849.00,  78,   0.17),
    ]

    for r, (name, unit_price, units, weight) in enumerate(products, 8):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=unit_price).number_format = '#,##0.00'
        revenue = round(unit_price * units, 2)
        ws.cell(row=r, column=3, value=revenue).number_format = '#,##0.00'
        ws.cell(row=r, column=4, value=weight).number_format = '0.00'

    # Row 23: blank separator

    # Row 24: Headers for summary statistics
    ws['A24'] = 'Summary Statistics'
    ws['A24'].font = Font(bold=True)

    # Row 25: Weighted average (C25) and annualized (D25)
    ws['A25'] = 'Weighted Avg / Annual'
    ws['C25'] = '=IFERROR(SUMPRODUCT(C8:C22,D8:D22)/SUM(C8:C22),0)'
    ws['D25'] = '=C25*12'

    ws['C25'].number_format = '#,##0.00'
    ws['D25'].number_format = '#,##0.00'

    # Rows 26–29: Projected cash flows (needed by NPV/IRR in row 30)
    ws['A26'] = 'CF Year 1'
    ws['A27'] = 'CF Year 2'
    ws['A28'] = 'CF Year 3'
    ws['A29'] = 'CF Year 4'

    # C26:C29 — projected cash flows derived from C25
    cash_flows = [
        '=C25*1.05',
        '=C25*1.10',
        '=C25*1.15',
        '=C25*1.20',
    ]
    for r, cf in enumerate(cash_flows, 26):
        cell = ws.cell(row=r, column=3, value=cf)
        cell.number_format = '#,##0.00'

    # Row 30: NPV, IRR, comparison
    ws['A30'] = 'NPV / IRR / Delta'
    ws['C30'] = '=NPV(0.10,C25:C29)'
    ws['D30'] = '=IRR(C25:C29)'
    ws['E30'] = '=C30-D30'

    ws['C30'].number_format = '#,##0.00'
    ws['D30'].number_format = '0.00%'
    ws['E30'].number_format = '#,##0.00'

    ws['A30'].font = Font(bold=True)
    ws['C30'].font = Font(bold=True)
    ws['D30'].font = Font(bold=True)
    ws['E30'].font = Font(bold=True)

    # ─── Sheet 2: Rates ────────────────────────────────────────────────────────
    ws2 = wb.create_sheet('Rates')
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 16

    ws2['A1'] = 'Currency'
    ws2['B1'] = 'Exchange Rate (to USD)'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)

    rates = [
        ('USD', 1.0000),
        ('EUR', 1.0842),
        ('GBP', 1.2719),
        ('JPY', 0.0067),
        ('CAD', 0.7381),
        ('AUD', 0.6524),
        ('CHF', 1.1203),
        ('CNY', 0.1382),
        ('INR', 0.0120),
        ('BRL', 0.1987),
    ]

    for r, (currency, rate) in enumerate(rates, 2):
        ws2.cell(row=r, column=1, value=currency)
        ws2.cell(row=r, column=2, value=rate).number_format = '0.0000'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print('Budget sheet: formula cells C5, D5, E5, C25, D25, C30, D30, E30 — NO comments')
    print('Rates sheet: currency lookup table for VLOOKUP')


create_initial()
