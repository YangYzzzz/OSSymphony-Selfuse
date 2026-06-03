"""
Initial Setup: Multi-criteria decision matrix for warehouse location selection
Task ID: calc_ops_076
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: SiteSelection ---
    ws = wb.active
    ws.title = 'SiteSelection'

    # Headers (Row 1)
    headers = ['Criterion', 'Weight', 'Site A', 'Site B', 'Site C', 'Site D']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_align

    # Criteria data (Rows 2-6) - exact values from task context
    data = [
        ['Lease Cost',              0.25, 8, 6, 9, 5],
        ['Proximity to Customers',  0.30, 7, 9, 5, 8],
        ['Labor Availability',      0.20, 6, 7, 8, 6],
        ['Infrastructure',          0.15, 9, 5, 7, 8],
        ['Tax Incentives',          0.10, 5, 8, 6, 9],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if c == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            if c == 2:
                cell.number_format = '0.00'

    # Row 7 is blank separator

    # Row 8: Weighted Score label only - C8:F8 left EMPTY (task requires adding SUMPRODUCT)
    ws.cell(row=8, column=1, value='Weighted Score')
    ws['A8'].font = Font(name='Calibri', size=11, bold=True)

    # Row 9: Rank label only - C9:F9 left EMPTY (task requires adding RANK)
    ws.cell(row=9, column=1, value='Rank')
    ws['A9'].font = Font(name='Calibri', size=11, bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 10
    for col_letter in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
