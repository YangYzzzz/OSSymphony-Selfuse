"""
Reward Script: Multi-criteria decision matrix for warehouse location selection
Task ID: calc_ops_076
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): SUMPRODUCT formulas in C8:F8 that compute weighted scores
  Component 2 (0.5): RANK formulas in C9:F9 that rank sites by weighted score
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_076'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: SiteSelection sheet must exist
    if 'SiteSelection' not in wb.sheetnames:
        print("FAIL: Sheet 'SiteSelection' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SiteSelection']

    # Component 1: SUMPRODUCT formulas in C8:F8 (0.5 points total, 0.125 each)
    # These cells should contain SUMPRODUCT formulas computing weighted scores
    # from the weight column (B2:B6) and respective site score columns.
    sumproduct_cols = {'C': 3, 'D': 4, 'E': 5, 'F': 6}
    for col_letter, col_idx in sumproduct_cols.items():
        try:
            cell_ref = f'{col_letter}8'
            val = ws.cell(row=8, column=col_idx).value
            if val is not None and isinstance(val, str) and 'SUMPRODUCT' in val.upper():
                print(f"PASS: Component 1 — {cell_ref} contains SUMPRODUCT formula: {val} (0.125 pts)")
                total_score += 0.125
            else:
                print(f"FAIL: Component 1 — {cell_ref} expected SUMPRODUCT formula, found: {repr(val)}")
        except Exception as e:
            print(f"ERROR: Component 1 — {col_letter}8: {e}")

    # Component 2: RANK formulas in C9:F9 (0.5 points total, 0.125 each)
    # These cells should contain RANK formulas that rank the sites by their
    # weighted scores in row 8.
    rank_cols = {'C': 3, 'D': 4, 'E': 5, 'F': 6}
    for col_letter, col_idx in rank_cols.items():
        try:
            cell_ref = f'{col_letter}9'
            val = ws.cell(row=9, column=col_idx).value
            if val is not None and isinstance(val, str) and 'RANK' in val.upper():
                print(f"PASS: Component 2 — {cell_ref} contains RANK formula: {val} (0.125 pts)")
                total_score += 0.125
            else:
                print(f"FAIL: Component 2 — {cell_ref} expected RANK formula, found: {repr(val)}")
        except Exception as e:
            print(f"ERROR: Component 2 — {col_letter}9: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
