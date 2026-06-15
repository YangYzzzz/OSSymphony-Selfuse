"""
Initial Setup: Project ROI Analysis - Pre-task state
Task ID: calc_fin_project_roi_061
Domain: libreoffice_calc

Creates a spreadsheet with 5 investment projects.
Columns A-D and H1 are filled. Columns E (ROI), F (Payback), G (NPV) are empty.
No formulas, no conditional formatting — those are the agent's task.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_project_roi_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProjectROI'

    # --- Row 1: Headers ---
    headers = [
        'Project',
        'Initial Investment',
        'Annual Cash Flow',
        'Project Life (Years)',
        'ROI',
        'Payback Period',
        'NPV'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        # Headers are NOT bold in initial (agent task includes making row 1 bold via formatting)

    # H1: discount rate
    ws['H1'] = 0.10
    ws['H1'].number_format = '0%'

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14

    # --- Rows 2-6: 5 realistic projects (A, B, C, D filled; E, F, G empty) ---
    # Projects designed so some pass (ROI>20%, payback<3yr) and some fail the criteria
    # ROI = (Annual*Life - Investment) / Investment
    # Payback = Investment / Annual
    # Project Alpha: invest=120000, annual=55000, life=4yr
    #   ROI = (55000*4 - 120000)/120000 = (220000-120000)/120000 = 83.3% PASS
    #   Payback = 120000/55000 = 2.18yr PASS
    # Project Beta: invest=250000, annual=60000, life=5yr
    #   ROI = (60000*5 - 250000)/250000 = 50000/250000 = 20% PASS (borderline)
    #   Payback = 250000/60000 = 4.17yr FAIL
    # Project Gamma: invest=80000, annual=15000, life=6yr
    #   ROI = (15000*6 - 80000)/80000 = 10000/80000 = 12.5% FAIL
    #   Payback = 80000/15000 = 5.33yr FAIL
    # Project Delta: invest=150000, annual=65000, life=3yr
    #   ROI = (65000*3 - 150000)/150000 = 45000/150000 = 30% PASS
    #   Payback = 150000/65000 = 2.31yr PASS
    # Project Epsilon: invest=300000, annual=45000, life=8yr
    #   ROI = (45000*8 - 300000)/300000 = 60000/300000 = 20% PASS (borderline)
    #   Payback = 300000/45000 = 6.67yr FAIL
    projects = [
        ('Project Alpha',   120000, 55000, 4),
        ('Project Beta',    250000, 60000, 5),
        ('Project Gamma',    80000, 15000, 6),
        ('Project Delta',   150000, 65000, 3),
        ('Project Epsilon', 300000, 45000, 8),
    ]

    currency_fmt = '$#,##0.00'

    for row_idx, (name, invest, cashflow, life) in enumerate(projects, 2):
        ws.cell(row=row_idx, column=1, value=name)

        cell_b = ws.cell(row=row_idx, column=2, value=invest)
        cell_b.number_format = currency_fmt

        cell_c = ws.cell(row=row_idx, column=3, value=cashflow)
        cell_c.number_format = currency_fmt

        ws.cell(row=row_idx, column=4, value=life)

        # Columns E, F, G: intentionally left EMPTY (agent must fill these)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
