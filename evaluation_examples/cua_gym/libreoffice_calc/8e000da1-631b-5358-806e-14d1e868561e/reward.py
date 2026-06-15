"""
Reward Script: Calculate ROI, Payback Period, and NPV for 5 projects with color-coded conditional formatting
Task ID: calc_fin_project_roi_061
Domain: libreoffice_calc
Scoring:
  Component 1: ROI formulas in E2:E6 with percentage format            - 0.25 pts
  Component 2: Payback period formulas in F2:F6 with decimal format    - 0.20 pts
  Component 3: NPV formulas in G2:G6 with currency format              - 0.25 pts
  Component 4: Row 1 headers bold (A1:G1)                             - 0.10 pts
  Component 5: Conditional formatting rules (green/red color-coding)   - 0.20 pts
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_project_roi_061'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: ProjectROI sheet must exist
    if 'ProjectROI' not in wb.sheetnames:
        print("FAIL: 'ProjectROI' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ProjectROI']

    # Component 1: ROI formulas in E2:E6 with percentage format (0.25 points)
    # Task requires =(Cx*Dx-Bx)/Bx formula with percentage format in E2:E6
    # This FAILS on initial (all None) and PASSES on golden
    try:
        roi_formula_count = 0
        roi_format_count = 0
        for row in range(2, 7):
            cell = ws.cell(row=row, column=5)
            val = cell.value
            # Check for ROI formula pattern: (C*D-B)/B structure
            if val and isinstance(val, str) and re.search(r'=\s*\(C\d\*D\d-B\d\)/B\d', val, re.IGNORECASE):
                roi_formula_count += 1
            # Check percentage number format
            if cell.number_format and '%' in cell.number_format:
                roi_format_count += 1

        if roi_formula_count == 5 and roi_format_count == 5:
            print(f"PASS: Component 1 — ROI formulas in E2:E6 with percentage format (0.25 pts)")
            total_score += 0.25
        elif roi_formula_count == 5:
            print(f"PASS: Component 1 (partial) — ROI formulas in E2:E6 but missing percentage format "
                  f"(formulas={roi_formula_count}/5, formats={roi_format_count}/5) (0.15 pts)")
            total_score += 0.15
        elif roi_formula_count > 0:
            partial = round(0.25 * roi_formula_count / 5, 2)
            print(f"FAIL: Component 1 (partial) — only {roi_formula_count}/5 ROI formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No ROI formulas found in E2:E6 "
                  f"(formula_count={roi_formula_count}, format_count={roi_format_count})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Payback period formulas in F2:F6 with decimal format (0.20 points)
    # Task requires =Bx/Cx formula with 1-decimal format in F2:F6
    # This FAILS on initial (all None) and PASSES on golden
    try:
        payback_formula_count = 0
        payback_format_count = 0
        for row in range(2, 7):
            cell = ws.cell(row=row, column=6)
            val = cell.value
            # Check for payback formula pattern: =B/C or =Bx/Cx
            if val and isinstance(val, str) and re.search(r'=\s*B\d\s*/\s*C\d', val, re.IGNORECASE):
                payback_formula_count += 1
            # Check decimal number format (e.g., "0.0")
            nf = cell.number_format
            if nf and nf not in ('General', '@', '') and ('0.' in nf or '.0' in nf) and '%' not in nf and '$' not in nf:
                payback_format_count += 1

        if payback_formula_count == 5 and payback_format_count == 5:
            print(f"PASS: Component 2 — Payback period formulas in F2:F6 with decimal format (0.20 pts)")
            total_score += 0.20
        elif payback_formula_count == 5:
            print(f"PASS: Component 2 (partial) — Payback formulas in F2:F6 but missing decimal format "
                  f"(formulas={payback_formula_count}/5, formats={payback_format_count}/5) (0.12 pts)")
            total_score += 0.12
        elif payback_formula_count > 0:
            partial = round(0.20 * payback_formula_count / 5, 2)
            print(f"FAIL: Component 2 (partial) — only {payback_formula_count}/5 payback formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No payback period formulas found in F2:F6 "
                  f"(formula_count={payback_formula_count}, format_count={payback_format_count})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: NPV formulas in G2:G6 with currency format (0.25 points)
    # Task requires PV-of-annuity formula =Cx*((1-(1+$H$1)^(-Dx))/$H$1)-Bx with currency format
    # This FAILS on initial (all None) and PASSES on golden
    try:
        npv_formula_count = 0
        npv_format_count = 0
        for row in range(2, 7):
            cell = ws.cell(row=row, column=7)
            val = cell.value
            # Check for NPV/PV formula - should reference H1 discount rate and have division
            if val and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                # Accept either NPV() function or PV-of-annuity formula pattern
                if (re.search(r'NPV\s*\(', val, re.IGNORECASE) or
                        (re.search(r'\$H\$1', val_upper) and re.search(r'B\d', val_upper) and re.search(r'C\d', val_upper))):
                    npv_formula_count += 1
            # Check currency number format
            nf = cell.number_format
            if nf and ('$' in nf or '€' in nf or '£' in nf or '#,##0' in nf):
                npv_format_count += 1

        if npv_formula_count == 5 and npv_format_count == 5:
            print(f"PASS: Component 3 — NPV formulas in G2:G6 with currency format (0.25 pts)")
            total_score += 0.25
        elif npv_formula_count == 5:
            print(f"PASS: Component 3 (partial) — NPV formulas in G2:G6 but missing currency format "
                  f"(formulas={npv_formula_count}/5, formats={npv_format_count}/5) (0.15 pts)")
            total_score += 0.15
        elif npv_formula_count > 0:
            partial = round(0.25 * npv_formula_count / 5, 2)
            print(f"FAIL: Component 3 (partial) — only {npv_formula_count}/5 NPV formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No NPV formulas found in G2:G6 "
                  f"(formula_count={npv_formula_count}, format_count={npv_format_count})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row 1 headers bold (0.10 points)
    # Task requires Row 1 to be bold
    # This FAILS on initial (not bold) and PASSES on golden (all bold)
    try:
        bold_count = 0
        for col in range(1, 8):  # A1:G1
            cell = ws.cell(row=1, column=col)
            if cell.font and cell.font.bold:
                bold_count += 1

        if bold_count == 7:
            print(f"PASS: Component 4 — Row 1 headers (A1:G1) are all bold (0.10 pts)")
            total_score += 0.10
        elif bold_count >= 4:
            print(f"PASS: Component 4 (partial) — {bold_count}/7 row 1 headers are bold (0.06 pts)")
            total_score += 0.06
        else:
            print(f"FAIL: Component 4 — Only {bold_count}/7 row 1 headers are bold")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting with green (passing) and red (failing) rules (0.20 points)
    # Task requires color-coding: green for ROI>20% AND payback<3yr, red otherwise
    # This FAILS on initial (no CF rules) and PASSES on golden (2 CF rules with correct colors/formulas)
    try:
        cf_rules_all = ws.conditional_formatting._cf_rules
        total_cf_rules = sum(len(rules) for rules in cf_rules_all.values())

        # Check for at least 2 CF rules
        if total_cf_rules < 2:
            print(f"FAIL: Component 5 — Only {total_cf_rules} conditional formatting rules found (need at least 2)")
        else:
            green_found = False
            red_found = False
            correct_range = False

            for cf_range, rules in cf_rules_all.items():
                range_str = str(cf_range)
                # Check that the CF applies to the data rows area
                if 'A2' in range_str or '2' in range_str:
                    correct_range = True

                for rule in rules:
                    # Check colors via dxf style
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            color_rgb = rule.dxf.fill.fgColor.rgb.upper()
                            # Green colors: include standard green shades used by Excel/LibreOffice
                            if any(g in color_rgb for g in ['00FF00', '92D050', '00B050', '70AD47', 'C6EFCE']):
                                green_found = True
                            # Red colors: include standard red shades
                            if any(r in color_rgb for r in ['FF0000', 'FF0000', 'FFC7CE', 'FF4444', 'C00000', 'FF4C4C']):
                                red_found = True
                        except Exception:
                            pass

                    # Also check by formula content if colors not readable
                    if hasattr(rule, 'formula') and rule.formula:
                        formula_str = str(rule.formula).upper()
                        # Green rule: both criteria met (ROI>=20% AND payback<=3)
                        if ('AND' in formula_str and 'E' in formula_str and 'F' in formula_str and
                                ('0.2' in formula_str or '20' in formula_str) and '3' in formula_str):
                            green_found = True
                        # Red rule: either criterion not met
                        if ('OR' in formula_str and 'E' in formula_str and 'F' in formula_str):
                            red_found = True

            if green_found and red_found and correct_range:
                print(f"PASS: Component 5 — Conditional formatting with green (passing) and red (failing) rules "
                      f"applied to data rows (0.20 pts)")
                total_score += 0.20
            elif (green_found or red_found) and correct_range:
                print(f"PASS: Component 5 (partial) — CF rules present but only one color found "
                      f"(green={green_found}, red={red_found}) (0.10 pts)")
                total_score += 0.10
            elif total_cf_rules >= 2:
                print(f"PASS: Component 5 (partial) — {total_cf_rules} CF rules present but colors not verified "
                      f"(green={green_found}, red={red_found}, correct_range={correct_range}) (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 5 — CF rules incomplete: total_rules={total_cf_rules}, "
                      f"green={green_found}, red={red_found}, correct_range={correct_range}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
