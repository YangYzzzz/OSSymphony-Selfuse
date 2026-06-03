"""
Initial Setup: Format sales performance dashboard
Task ID: calc_gsd_023
Domain: libreoffice_calc

Creates a sales performance spreadsheet with 60 reps' data.
No freeze panes, no currency formatting, no conditional formatting.
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reps"

    # Headers
    headers = ["Rep ID", "Name", "Region", "Calls Made", "Deals Closed",
               "Revenue", "Target", "Variance"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic names
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei",
        "Carlos", "Aisha", "Robert", "Yuki", "Michael", "Fatima", "Andrew",
        "Sophia", "Daniel", "Liam", "Nadia", "Ethan", "Grace",
        "Oliver", "Zara", "Noah", "Hannah", "Leo", "Amara", "Ryan",
        "Isabelle", "Kevin", "Chloe", "Nathan", "Ava", "Brandon", "Layla",
        "Tyler", "Mia", "Jason", "Emma", "Patrick", "Luna",
        "Connor", "Jade", "Derek", "Samira", "Kyle", "Tanya", "Alex",
        "Nina", "Chris", "Alicia", "Ben", "Victoria", "Sam", "Rachel",
        "Ian", "Olivia", "Mark", "Diana", "Jordan", "Tessa"
    ]
    last_names = [
        "Chen", "Johnson", "Petrov", "Williams", "Sharma", "Kim", "Zhang",
        "Garcia", "Hassan", "Taylor", "Tanaka", "Brown", "Ali", "Morrison",
        "Reyes", "Park", "O'Brien", "Gupta", "Foster", "Nakamura",
        "Davis", "Okafor", "Miller", "Lee", "Santos", "Diallo", "Clark",
        "Moreau", "Wright", "Patel", "Hughes", "Rossi", "Evans", "Muller",
        "Adams", "Nguyen", "Scott", "Kowalski", "Collins", "Dubois",
        "Reed", "Johansson", "Grant", "Osman", "Harper", "Ivanova", "Cross",
        "Ferreira", "Bell", "Torres", "Ward", "Russo", "Hunt", "Lambert",
        "Cooper", "Cho", "Young", "Alves", "Phillips", "Bergstrom"
    ]
    regions = ["Northeast", "Southeast", "Midwest", "West Coast",
               "Southwest", "Pacific Northwest", "Mid-Atlantic", "Central"]

    for i in range(60):
        row = i + 2
        rep_id = f"SR-{1001 + i}"
        name = f"{first_names[i]} {last_names[i]}"
        region = regions[i % len(regions)]
        calls = random.randint(80, 350)
        deals = random.randint(3, 35)
        revenue = random.randint(12000, 285000)
        target = random.randint(50000, 200000)
        variance = revenue - target

        ws.cell(row=row, column=1, value=rep_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=region)
        ws.cell(row=row, column=4, value=calls)
        ws.cell(row=row, column=5, value=deals)
        ws.cell(row=row, column=6, value=revenue)
        ws.cell(row=row, column=7, value=target)
        ws.cell(row=row, column=8, value=variance)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

    # NO freeze panes (task asks agent to freeze)
    # NO currency formatting (task asks agent to format)
    # NO conditional formatting (task asks agent to add)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
