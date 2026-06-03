"""
Reward Script: Format sales performance dashboard
Task ID: calc_gsd_023
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): Freeze panes set to A2 (header row frozen)
  - Component 2 (0.35): Revenue column (F2:F61) has currency format with red negatives
  - Component 3 (0.30): Conditional formatting on F2:F61 for top-10 highlight with green background
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Reps' sheet must exist
    if 'Reps' not in wb.sheetnames:
        print("FAIL: 'Reps' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Reps']

    # Component 1: Freeze panes set to A2 (0.35 points)
    # Initial has freeze_panes=None; golden has freeze_panes='A2'
    try:
        freeze = ws.freeze_panes
        if freeze == 'A2':
            print(f"PASS: Component 1 — freeze_panes is 'A2' (row 1 frozen) (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — expected freeze_panes='A2', found: {freeze}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Revenue column (F2:F61) has currency format with red negatives (0.35 points)
    # Initial has 'General'; golden has '$#,##0.00;[Red]-$#,##0.00'
    # We check a sample of cells and require most to have a currency-like format
    try:
        currency_count = 0
        total_checked = 0
        sample_rows = list(range(2, 62))  # F2:F61
        for r in sample_rows:
            cell = ws.cell(row=r, column=6)
            nf = cell.number_format
            total_checked += 1
            # Accept any currency format that includes '$' and '[Red]' for negative display
            if nf and '$' in str(nf) and '[Red]' in str(nf):
                currency_count += 1

        ratio = currency_count / total_checked if total_checked > 0 else 0
        if ratio >= 0.9:
            print(f"PASS: Component 2 — {currency_count}/{total_checked} F cells have currency format with red negatives (0.35 pts)")
            total_score += 0.35
        elif ratio >= 0.5:
            partial = round(0.35 * ratio, 2)
            print(f"PARTIAL: Component 2 — {currency_count}/{total_checked} F cells have currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — only {currency_count}/{total_checked} F cells have currency format with [Red] negatives")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on F2:F61 for top-10 with green background (0.30 points)
    # Initial has no conditional formatting; golden has expression-based top-10 rule with green fill
    try:
        cf_list = list(ws.conditional_formatting)
        found_top10_rule = False

        for cf in cf_list:
            cf_range = str(cf)
            # Check if range covers F2:F61 (may appear as "F2:F61" in the range string)
            if 'F' not in cf_range:
                continue

            for rule in cf.rules:
                # The golden uses expression type with RANK formula for top-10
                # Accept either: expression with RANK<=10, or a top10 type rule
                is_top10 = False

                if rule.type == 'top10' and rule.rank == 10:
                    is_top10 = True
                elif rule.type == 'expression':
                    formulas = rule.formula if rule.formula else []
                    for f in formulas:
                        f_upper = str(f).upper().replace(' ', '')
                        if 'RANK' in f_upper and '<=10' in f_upper:
                            is_top10 = True
                            break

                if is_top10:
                    # Check for green-ish fill
                    has_green = False
                    if rule.dxf and rule.dxf.fill:
                        fg = rule.dxf.fill.fgColor
                        if fg and hasattr(fg, 'rgb') and fg.rgb:
                            rgb_str = str(fg.rgb).upper()
                            # Check green channel is dominant: FF00FF00 or similar greens
                            # Green: G component high, R and B components low
                            if len(rgb_str) == 8:
                                r_val = int(rgb_str[2:4], 16)
                                g_val = int(rgb_str[4:6], 16)
                                b_val = int(rgb_str[6:8], 16)
                                if g_val > 128 and g_val > r_val:
                                    has_green = True

                    if has_green:
                        print(f"PASS: Component 3 — Top-10 conditional formatting on F column with green fill (0.30 pts)")
                        total_score += 0.30
                        found_top10_rule = True
                        break
                    else:
                        # Top-10 rule exists but without clear green fill — partial credit
                        print(f"PARTIAL: Component 3 — Top-10 rule found but fill color not clearly green (0.15 pts)")
                        total_score += 0.15
                        found_top10_rule = True
                        break

            if found_top10_rule:
                break

        if not found_top10_rule:
            print(f"FAIL: Component 3 — No top-10 conditional formatting rule found on F column")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
