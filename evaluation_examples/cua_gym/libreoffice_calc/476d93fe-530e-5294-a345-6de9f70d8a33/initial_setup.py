"""
Initial Setup: Software License Compliance Tracker
Task ID: calc_grs_092
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: License Inventory ==========
    ws1 = wb.active
    ws1.title = "License Inventory"

    headers = [
        "Software Name", "Vendor", "License Type", "Licenses Purchased",
        "Licenses Deployed", "Compliance Status", "Annual Cost",
        "Renewal Date", "Contract Owner", "Notes"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Realistic license data - mix of compliant and non-compliant
    # Compliance Status column (F) is intentionally LEFT BLANK - task asks agent to add formula
    data = [
        ["Microsoft Office 365", "Microsoft", "Subscription", 250, 243, "", 89750, "2026-06-15", "Diana Mitchell", "Enterprise E3 plan"],
        ["Adobe Creative Cloud", "Adobe", "Subscription", 45, 52, "", 32400, "2026-03-01", "Robert Zhang", "Design team + marketing overflow"],
        ["Salesforce CRM", "Salesforce", "Per User", 120, 118, "", 172800, "2026-09-30", "Jennifer Park", "Sales and support teams"],
        ["AutoCAD", "Autodesk", "Per Device", 30, 35, "", 52500, "2026-04-20", "Thomas Rivera", "Engineering dept needs more seats"],
        ["Slack Business+", "Slack Technologies", "Per User", 300, 287, "", 45600, "2026-08-12", "Sarah Chen", "Company-wide communication"],
        ["Tableau Desktop", "Salesforce", "Per User", 25, 22, "", 42500, "2026-11-01", "Marcus Johnson", "Analytics team"],
        ["Jira Software", "Atlassian", "Per User", 150, 148, "", 15000, "2026-07-15", "Aisha Patel", "Dev and PM teams"],
        ["Zoom Enterprise", "Zoom", "Site License", 1, 1, "", 28000, "2026-05-28", "Diana Mitchell", "Unlimited users under site license"],
        ["GitHub Enterprise", "Microsoft", "Per User", 80, 94, "", 31680, "2026-10-10", "Thomas Rivera", "Dev team grew after Q3 hiring"],
        ["Figma Organization", "Figma", "Per User", 40, 38, "", 18000, "2026-12-01", "Robert Zhang", "Design and product teams"],
        ["Datadog Pro", "Datadog", "Per Device", 60, 42, "", 108000, "2027-01-15", "Aisha Patel", "Monitoring - some hosts decommissioned"],
        ["Okta Identity", "Okta", "Per User", 350, 310, "", 52500, "2026-02-28", "Sarah Chen", "SSO for all employees"],
        ["Snowflake", "Snowflake", "Subscription", 5, 5, "", 96000, "2026-08-01", "Marcus Johnson", "Data warehouse - credit based"],
        ["Postman Enterprise", "Postman", "Per User", 50, 28, "", 14400, "2026-06-30", "Jennifer Park", "API testing - many inactive"],
        ["Miro Business", "Miro", "Per User", 100, 45, "", 16000, "2026-11-15", "Diana Mitchell", "Collaboration boards - low adoption"],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 7:  # Annual Cost column - currency format
                cell.number_format = '$#,##0.00'
            if c == 8 and val:  # Renewal Date
                cell.number_format = 'yyyy-mm-dd'

    # Set column widths
    col_widths = {'A': 24, 'B': 20, 'C': 16, 'D': 18, 'E': 18,
                  'F': 20, 'G': 16, 'H': 14, 'I': 18, 'J': 35}
    for col_letter, width in col_widths.items():
        ws1.column_dimensions[col_letter].width = width

    # Freeze header row
    ws1.freeze_panes = "A2"

    # ========== Sheet 2: Summary (headers only, no formulas) ==========
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "License Compliance Summary"
    ws2["A1"].font = Font(size=14, bold=True)

    # Just labels, no values - task asks agent to add these
    ws2["A3"] = "Total Annual License Cost:"
    ws2["A4"] = "Number of Non-Compliant Software:"
    ws2["A5"] = "Potential License Risk Value:"
    ws2["A3"].font = Font(bold=True)
    ws2["A4"].font = Font(bold=True)
    ws2["A5"].font = Font(bold=True)
    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 20

    # ========== Sheet 3: Utilization Analysis (headers only) ==========
    ws3 = wb.create_sheet("Utilization Analysis")
    util_headers = ["Software Name", "Licenses Purchased", "Licenses Deployed",
                    "Usage Percentage", "Status"]
    for col, h in enumerate(util_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    ws3.column_dimensions['A'].width = 24
    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['C'].width = 20
    ws3.column_dimensions['D'].width = 18
    ws3.column_dimensions['E'].width = 22

    # ========== Sheet 4: Renewal Calendar (empty) ==========
    ws4 = wb.create_sheet("Renewal Calendar")
    ws4["A1"] = "Renewal Calendar"
    ws4["A1"].font = Font(size=14, bold=True)
    ws4.column_dimensions['A'].width = 16
    ws4.column_dimensions['B'].width = 24
    ws4.column_dimensions['C'].width = 16
    ws4.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
