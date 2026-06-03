"""
Reward Script: Software License Compliance Tracker
Task ID: calc_grs_092
Domain: libreoffice_calc
Scoring:
  Component 1: Compliance formulas in F column (0.20)
  Component 2: Conditional formatting for NON-COMPLIANT (0.15)
  Component 3: Data validation dropdown for License Type (0.10)
  Component 4: Summary formulas in B3:B5 (0.20)
  Component 5: Utilization Analysis data rows with formulas (0.15)
  Component 6: Utilization Analysis conditional formatting for underutilized (0.05)
  Component 7: Renewal Calendar populated with sorted entries (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_092'


def persist_app_state(domain):
    """Best-effort save via Ctrl+S for any open LibreOffice instance."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check required sheets exist as precondition
    required_sheets = ['License Inventory', 'Summary', 'Utilization Analysis', 'Renewal Calendar']
    for sn in required_sheets:
        if sn not in wb.sheetnames:
            # Sheets exist in initial too, but we need them for checking task changes
            pass  # Don't gate on this since initial also has them

    # ---------------------------------------------------------------
    # Component 1: Compliance Status formulas in F2:F16 (0.20 points)
    # Initial: F2:F16 are empty. Golden: each has =IF(E<=D,"Compliant","NON-COMPLIANT")
    # ---------------------------------------------------------------
    try:
        ws = wb['License Inventory']
        formula_count = 0
        correct_formula_count = 0
        for row in range(2, 17):  # F2:F16 = 15 rows
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_count += 1
                # Check it's an IF formula referencing deployed vs purchased
                upper_val = val.upper().replace(" ", "")
                if 'IF(' in upper_val and 'COMPLIANT' in upper_val:
                    correct_formula_count += 1

        if correct_formula_count >= 13:  # Allow small tolerance (13/15)
            print(f"PASS: Component 1 — {correct_formula_count}/15 compliance formulas found (0.20 pts)")
            total_score += 0.20
        elif correct_formula_count >= 8:
            partial = 0.10
            print(f"PARTIAL: Component 1 — {correct_formula_count}/15 compliance formulas (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correct_formula_count}/15 compliance formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Conditional formatting for NON-COMPLIANT in red/bold (0.15 points)
    # Initial: no CF rules. Golden: CF on F2:F16 cellIs "NON-COMPLIANT" with red fill + bold
    # ---------------------------------------------------------------
    try:
        ws = wb['License Inventory']
        cf_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                # Look for a rule that targets NON-COMPLIANT text
                has_non_compliant = False
                if hasattr(rule, 'formula') and rule.formula:
                    for f in rule.formula:
                        if 'NON-COMPLIANT' in str(f).upper() or 'NON_COMPLIANT' in str(f).upper():
                            has_non_compliant = True
                if rule.type == 'cellIs' and has_non_compliant:
                    cf_found = True
                    # Check if styling includes red or bold
                    has_red_or_bold = False
                    if hasattr(rule, 'dxf') and rule.dxf:
                        if rule.dxf.font and rule.dxf.font.bold:
                            has_red_or_bold = True
                        if rule.dxf.fill and rule.dxf.fill.fgColor:
                            rgb = str(rule.dxf.fill.fgColor.rgb).upper()
                            if 'FF0000' in rgb:
                                has_red_or_bold = True
                    if has_red_or_bold:
                        print(f"PASS: Component 2 — CF for NON-COMPLIANT with red/bold styling (0.15 pts)")
                        total_score += 0.15
                    else:
                        print(f"PARTIAL: Component 2 — CF for NON-COMPLIANT found but styling incomplete (0.08 pts)")
                        total_score += 0.08
                    break
            if cf_found:
                break

        if not cf_found:
            # Also accept formula-based CF that references NON-COMPLIANT
            for cf in ws.conditional_formatting:
                for rule in cf.rules:
                    if rule.type == 'expression' and hasattr(rule, 'formula') and rule.formula:
                        for f in rule.formula:
                            if 'NON-COMPLIANT' in str(f).upper() or 'NON_COMPLIANT' in str(f).upper():
                                cf_found = True
                                print(f"PASS: Component 2 — Formula-based CF for NON-COMPLIANT found (0.15 pts)")
                                total_score += 0.15
                                break
                    if cf_found:
                        break
                if cf_found:
                    break

        if not cf_found:
            print("FAIL: Component 2 — No conditional formatting for NON-COMPLIANT found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Data validation dropdown for License Type C2:C16 (0.10 points)
    # Initial: no DV. Golden: list DV with license type options
    # ---------------------------------------------------------------
    try:
        ws = wb['License Inventory']
        dv_found = False
        for dv in ws.data_validations.dataValidation:
            if dv.type == 'list' and dv.formula1:
                formula_str = str(dv.formula1).upper()
                # Check it includes at least some of the expected license types
                expected_types = ['PERPETUAL', 'SUBSCRIPTION', 'PER USER', 'PER DEVICE', 'SITE LICENSE', 'OPEN SOURCE']
                matches = sum(1 for t in expected_types if t in formula_str)
                if matches >= 4:  # At least 4 of 6 types present
                    dv_found = True
                    print(f"PASS: Component 3 — License Type dropdown with {matches}/6 expected types (0.10 pts)")
                    total_score += 0.10
                    break
        if not dv_found:
            print("FAIL: Component 3 — No data validation dropdown for License Type found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Summary formulas in B3:B5 (0.20 points)
    # Initial: B3:B5 are empty. Golden: SUM, COUNTIF, SUMPRODUCT formulas
    # ---------------------------------------------------------------
    try:
        ws_sum = wb['Summary']
        summary_score = 0.0

        # B3: Total Annual License Cost (SUM formula)
        b3 = ws_sum['B3'].value
        if b3 and isinstance(b3, str) and '=' in b3:
            upper_b3 = b3.upper().replace(" ", "")
            if 'SUM' in upper_b3:
                summary_score += 0.07
                print(f"PASS: Component 4a — SUM formula for total cost in B3: {b3}")
            else:
                print(f"FAIL: Component 4a — Expected SUM formula in B3, found: {b3}")
        else:
            print(f"FAIL: Component 4a — B3 is empty or not a formula: {b3}")

        # B4: Number of Non-Compliant (COUNTIF formula)
        b4 = ws_sum['B4'].value
        if b4 and isinstance(b4, str) and '=' in b4:
            upper_b4 = b4.upper().replace(" ", "")
            if 'COUNTIF' in upper_b4 and 'COMPLIANT' in upper_b4:
                summary_score += 0.07
                print(f"PASS: Component 4b — COUNTIF formula for non-compliant count in B4: {b4}")
            else:
                print(f"FAIL: Component 4b — Expected COUNTIF formula in B4, found: {b4}")
        else:
            print(f"FAIL: Component 4b — B4 is empty or not a formula: {b4}")

        # B5: Potential License Risk Value (SUMPRODUCT or similar)
        b5 = ws_sum['B5'].value
        if b5 and isinstance(b5, str) and '=' in b5:
            upper_b5 = b5.upper().replace(" ", "")
            if ('SUMPRODUCT' in upper_b5 or 'SUMIF' in upper_b5) and 'COMPLIANT' in upper_b5:
                summary_score += 0.06
                print(f"PASS: Component 4c — Risk value formula in B5: {b5}")
            else:
                # Accept any formula that references compliance
                if 'COMPLIANT' in upper_b5 or 'SUM' in upper_b5:
                    summary_score += 0.04
                    print(f"PARTIAL: Component 4c — Formula in B5 partially matches: {b5}")
                else:
                    print(f"FAIL: Component 4c — Expected risk value formula in B5, found: {b5}")
        else:
            print(f"FAIL: Component 4c — B5 is empty or not a formula: {b5}")

        total_score += summary_score
        print(f"  Component 4 total: {summary_score:.2f}/0.20 pts")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Utilization Analysis data rows with formulas (0.15 points)
    # Initial: only header row. Golden: 15 data rows (rows 2-16) with usage % and status formulas
    # ---------------------------------------------------------------
    try:
        ws_util = wb['Utilization Analysis']
        data_rows = 0
        usage_formula_count = 0
        status_formula_count = 0

        for row in range(2, ws_util.max_row + 1):
            a_val = ws_util.cell(row=row, column=1).value
            if a_val is not None:
                data_rows += 1
            d_val = ws_util.cell(row=row, column=4).value
            if d_val and isinstance(d_val, str) and '=' in d_val:
                usage_formula_count += 1
            e_val = ws_util.cell(row=row, column=5).value
            if e_val and isinstance(e_val, str) and '=' in e_val:
                status_formula_count += 1

        if data_rows >= 10 and usage_formula_count >= 10 and status_formula_count >= 10:
            print(f"PASS: Component 5 — Util Analysis has {data_rows} data rows, {usage_formula_count} usage formulas, {status_formula_count} status formulas (0.15 pts)")
            total_score += 0.15
        elif data_rows >= 5:
            partial = 0.08
            print(f"PARTIAL: Component 5 — {data_rows} rows, {usage_formula_count} usage formulas, {status_formula_count} status formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {data_rows} data rows in Utilization Analysis")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Utilization Analysis conditional formatting for underutilized (0.05 points)
    # Initial: no CF. Golden: CF highlighting rows where usage < 50%
    # ---------------------------------------------------------------
    try:
        ws_util = wb['Utilization Analysis']
        util_cf_found = False
        for cf in ws_util.conditional_formatting:
            for rule in cf.rules:
                if hasattr(rule, 'formula') and rule.formula:
                    for f in rule.formula:
                        f_str = str(f).upper().replace(" ", "")
                        if '0.5' in f_str or '50' in f_str or '<' in f_str:
                            util_cf_found = True
                            break
                if util_cf_found:
                    break
            if util_cf_found:
                break

        if util_cf_found:
            print(f"PASS: Component 6 — Underutilized CF found in Utilization Analysis (0.05 pts)")
            total_score += 0.05
        else:
            print("FAIL: Component 6 — No underutilized highlighting CF in Utilization Analysis")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # ---------------------------------------------------------------
    # Component 7: Renewal Calendar populated with sorted entries (0.15 points)
    # Initial: only "Renewal Calendar" header in A1. Golden: full calendar with 13+ renewal entries
    # ---------------------------------------------------------------
    try:
        ws_cal = wb['Renewal Calendar']
        # Count non-empty rows beyond row 1 (header area)
        cal_data_rows = 0
        has_date_data = False
        for row in range(2, ws_cal.max_row + 1):
            row_vals = [ws_cal.cell(row=row, column=c).value for c in range(1, ws_cal.max_column + 1)]
            if any(v is not None for v in row_vals):
                cal_data_rows += 1
                # Check if any cell looks like a date or software name
                for v in row_vals:
                    if v and isinstance(v, str) and ('2026' in str(v) or '2027' in str(v)):
                        has_date_data = True

        if cal_data_rows >= 10 and has_date_data:
            print(f"PASS: Component 7 — Renewal Calendar has {cal_data_rows} data rows with date info (0.15 pts)")
            total_score += 0.15
        elif cal_data_rows >= 5:
            partial = 0.08
            print(f"PARTIAL: Component 7 — {cal_data_rows} data rows in Renewal Calendar ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 — Only {cal_data_rows} data rows in Renewal Calendar (need >= 10)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
