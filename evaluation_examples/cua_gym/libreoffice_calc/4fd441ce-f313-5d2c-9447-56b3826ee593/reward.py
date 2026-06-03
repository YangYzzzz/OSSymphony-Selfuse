"""
Reward Script: YTD Summary with cross-sheet references, cumulative formulas, formatting, and conditional formatting
Task ID: calc_gsd_033
Domain: libreoffice_calc
Scoring:
  Component 1: Revenue formulas in B2:B13 referencing monthly sheets (0.25)
  Component 2: Cumulative Revenue formulas in C2:C13 (0.20)
  Component 3: Percentage formulas in D2:D13 (0.15)
  Component 4: Bold header row 1 (0.10)
  Component 5: Borders on A1:D13 (0.10)
  Component 6: Currency number format on B2:C13 and percentage format on D2:D13 (0.10)
  Component 7: Conditional formatting on D2:D13 (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_033'

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'YTD Summary' not in wb.sheetnames:
        print("CRITICAL: 'YTD Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['YTD Summary']

    # Component 1: Revenue formulas in B2:B13 referencing monthly sheets (0.25 points)
    # Each cell should reference the corresponding month sheet's B2, e.g., ='Jan'!B2
    try:
        ref_count = 0
        for i, month in enumerate(MONTHS):
            row = i + 2
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str):
                # Accept various formula forms: ='Jan'!B2, =Jan!B2, =Jan.B2, ='Jan'.B2
                val_upper = val.upper().replace(" ", "").replace("'", "").replace(".", "!")
                expected_upper = f"={month.upper()}!B2"
                if expected_upper in val_upper:
                    ref_count += 1
        if ref_count == 12:
            print(f"PASS: Component 1 — All 12 Revenue formulas reference monthly sheets (0.25 pts)")
            total_score += 0.25
        elif ref_count >= 6:
            partial = round(0.25 * ref_count / 12, 3)
            print(f"PARTIAL: Component 1 — {ref_count}/12 Revenue formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {ref_count}/12 Revenue formulas reference monthly sheets")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cumulative Revenue formulas in C2:C13 (0.20 points)
    # C2 should be =B2, C3 should be =SUM($B$2:B3) or similar cumulative pattern
    try:
        cum_count = 0
        for i in range(12):
            row = i + 2
            val = ws.cell(row=row, column=3).value
            if val is not None and isinstance(val, str):
                val_clean = val.upper().replace(" ", "")
                if row == 2:
                    # C2: could be =B2 or =SUM($B$2:B2) or =SUM(B2:B2)
                    if "B2" in val_clean and val_clean.startswith("="):
                        cum_count += 1
                else:
                    # C3-C13: should be cumulative SUM involving B column
                    # Accept =SUM($B$2:B3), =SUM(B$2:B3), =B2+B3, etc.
                    if "SUM" in val_clean or ("+" in val_clean and "B" in val_clean):
                        # Verify it references the current row's B cell
                        if f"B{row}" in val_clean:
                            cum_count += 1
        if cum_count == 12:
            print(f"PASS: Component 2 — All 12 Cumulative Revenue formulas correct (0.20 pts)")
            total_score += 0.20
        elif cum_count >= 6:
            partial = round(0.20 * cum_count / 12, 3)
            print(f"PARTIAL: Component 2 — {cum_count}/12 Cumulative formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {cum_count}/12 Cumulative Revenue formulas correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage of annual target formulas in D2:D13 (0.15 points)
    # Each should be =C<row>/$D$1 or similar
    try:
        pct_count = 0
        for i in range(12):
            row = i + 2
            val = ws.cell(row=row, column=4).value
            if val is not None and isinstance(val, str):
                val_clean = val.upper().replace(" ", "")
                if val_clean.startswith("="):
                    # Should reference C<row> and D1 (the target)
                    has_c_ref = f"C{row}" in val_clean
                    has_d1_ref = "D1" in val_clean or "D$1" in val_clean or "$D$1" in val_clean
                    # Also accept division by 6000000 directly
                    has_target = has_d1_ref or "6000000" in val_clean
                    if has_c_ref and has_target:
                        pct_count += 1
        if pct_count == 12:
            print(f"PASS: Component 3 — All 12 Percentage formulas correct (0.15 pts)")
            total_score += 0.15
        elif pct_count >= 6:
            partial = round(0.15 * pct_count / 12, 3)
            print(f"PARTIAL: Component 3 — {pct_count}/12 Percentage formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {pct_count}/12 Percentage formulas correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Bold header row 1 (0.10 points)
    # All 4 cells in row 1 (A1:D1) should be bold
    try:
        bold_count = 0
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            if cell.font.bold:
                bold_count += 1
        if bold_count == 4:
            print(f"PASS: Component 4 — All 4 header cells are bold (0.10 pts)")
            total_score += 0.10
        elif bold_count >= 2:
            partial = round(0.10 * bold_count / 4, 3)
            print(f"PARTIAL: Component 4 — {bold_count}/4 header cells bold ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {bold_count}/4 header cells are bold")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Borders on A1:D13 (0.10 points)
    # All cells in A1:D13 should have thin borders on all sides
    try:
        bordered_count = 0
        total_cells = 13 * 4  # 13 rows, 4 columns
        for row in range(1, 14):
            for col in range(1, 5):
                cell = ws.cell(row=row, column=col)
                b = cell.border
                if (b.left.style is not None and
                    b.right.style is not None and
                    b.top.style is not None and
                    b.bottom.style is not None):
                    bordered_count += 1
        if bordered_count == total_cells:
            print(f"PASS: Component 5 — All {total_cells} cells have borders (0.10 pts)")
            total_score += 0.10
        elif bordered_count >= total_cells * 0.5:
            partial = round(0.10 * bordered_count / total_cells, 3)
            print(f"PARTIAL: Component 5 — {bordered_count}/{total_cells} cells have borders ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {bordered_count}/{total_cells} cells have borders")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Number formats - currency on B2:C13, percentage on D2:D13 (0.10 points)
    try:
        fmt_ok = 0
        fmt_total = 36  # 12 rows * 3 columns (B, C, D)

        for row in range(2, 14):
            # B column: USD currency with 0 decimals
            b_fmt = ws.cell(row=row, column=2).number_format
            if '$' in b_fmt and '0' in b_fmt:
                fmt_ok += 1

            # C column: USD currency with 0 decimals
            c_fmt = ws.cell(row=row, column=3).number_format
            if '$' in c_fmt and '0' in c_fmt:
                fmt_ok += 1

            # D column: percentage with 1 decimal
            d_fmt = ws.cell(row=row, column=4).number_format
            if '%' in d_fmt:
                fmt_ok += 1

        if fmt_ok == fmt_total:
            print(f"PASS: Component 6 — All number formats correct (0.10 pts)")
            total_score += 0.10
        elif fmt_ok >= fmt_total * 0.5:
            partial = round(0.10 * fmt_ok / fmt_total, 3)
            print(f"PARTIAL: Component 6 — {fmt_ok}/{fmt_total} format checks pass ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Only {fmt_ok}/{fmt_total} format checks pass")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Conditional formatting on D2:D13 (0.10 points)
    # Should have rules for >=100% green, >=80% yellow, <80% red
    try:
        cf_rules = list(ws.conditional_formatting)
        # Count distinct color rules found on D column
        green_colors = {'FF00B050', 'FF00FF00', '0000FF00', '0000B050'}
        yellow_colors = {'FFFFFF00', '00FFFF00'}
        red_colors = {'FFFF0000', '00FF0000'}
        found_colors = set()  # track which color categories are present

        for cf in cf_rules:
            range_str = str(cf).upper()
            # Check if the range covers D column in rows 2-13
            if 'D' in range_str:
                for rule in cf.rules:
                    fill_rgb = None
                    if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                        try:
                            fill_rgb = rule.dxf.fill.fgColor.rgb
                        except Exception:
                            pass

                    if fill_rgb:
                        if fill_rgb in green_colors:
                            found_colors.add('green')
                        if fill_rgb in yellow_colors:
                            found_colors.add('yellow')
                        if fill_rgb in red_colors:
                            found_colors.add('red')

        cf_count = len(found_colors)
        if cf_count == 3:
            print(f"PASS: Component 7 — All 3 conditional formatting rules found (0.10 pts)")
            total_score += 0.10
        elif cf_count >= 1:
            partial = round(0.10 * cf_count / 3, 3)
            print(f"PARTIAL: Component 7 — {cf_count}/3 conditional formatting rules found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 — No conditional formatting rules found on D2:D13")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification (LibreOffice may have unsaved changes)
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
