"""
Initial Setup: Create annual tracker spreadsheet with 12 monthly sheets and a YTD Summary sheet.
Task ID: calc_gsd_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Monthly revenue values (realistic tech company monthly revenue)
    monthly_revenues = {
        'Jan': 423500,
        'Feb': 398200,
        'Mar': 512800,
        'Apr': 487600,
        'May': 534100,
        'Jun': 561300,
        'Jul': 478900,
        'Aug': 502400,
        'Sep': 548700,
        'Oct': 589200,
        'Nov': 621500,
        'Dec': 673800,
    }

    # Realistic breakdown categories for monthly sheets
    categories = [
        'Software Licensing',
        'Cloud Services',
        'Consulting',
        'Maintenance & Support',
        'Training Services',
        'Custom Development',
        'Hardware Sales',
    ]

    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # Create monthly sheets
    for i, month in enumerate(month_names):
        if i == 0:
            ws = wb.active
            ws.title = month
        else:
            ws = wb.create_sheet(month)

        total_rev = monthly_revenues[month]

        # Headers
        ws['A1'] = 'Category'
        ws['B1'] = 'Amount'
        ws['A1'].font = Font(bold=True)
        ws['B1'].font = Font(bold=True)

        # Total Revenue in B2 (this is what YTD Summary will reference)
        ws['A2'] = 'Total Revenue'
        ws['B2'] = total_rev

        # Breakdown rows (for realism, these don't need to sum to total exactly)
        import random
        random.seed(42 + i)
        portions = [random.uniform(0.05, 0.25) for _ in categories]
        total_portions = sum(portions)
        for j, cat in enumerate(categories):
            row = j + 3
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=round(total_rev * portions[j] / total_portions, 2))

        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 18

    # Create YTD Summary sheet
    ws_ytd = wb.create_sheet('YTD Summary')

    # Headers in row 1
    ws_ytd['A1'] = 'Month'
    ws_ytd['B1'] = 'Revenue'
    ws_ytd['C1'] = 'Cumulative Revenue'
    ws_ytd['D1'] = 6000000  # Annual target value

    # Month names in A2:A13
    for idx, month in enumerate(month_names):
        ws_ytd.cell(row=idx + 2, column=1, value=month)

    # B2:B13, C2:C13, D2:D13 intentionally left EMPTY
    # The task is for the agent to fill these with formulas

    # Column widths
    ws_ytd.column_dimensions['A'].width = 12
    ws_ytd.column_dimensions['B'].width = 18
    ws_ytd.column_dimensions['C'].width = 22
    ws_ytd.column_dimensions['D'].width = 22

    # NO formatting, NO borders, NO bold on headers, NO conditional formatting
    # These are all part of the task the agent must complete

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
