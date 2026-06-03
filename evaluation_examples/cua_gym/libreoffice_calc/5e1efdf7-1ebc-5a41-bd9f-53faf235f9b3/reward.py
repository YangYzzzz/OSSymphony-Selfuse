"""
Reward Script: Enable sheet-level protection on 'Parameters' sheet with password,
               keeping B2:B10 editable and all other cells protected.
Task ID: calc_gg3_034
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Sheet protection enabled on 'Parameters'
  Component 2 (0.2): Password matches 'readonly123'
  Component 3 (0.3): Cells B2:B10 are unlocked (editable)
  Component 4 (0.2): Other cells remain locked (protected)
"""

import os
import openpyxl
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_034'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Parameters' sheet must exist
    if 'Parameters' not in wb.sheetnames:
        print("FAIL: 'Parameters' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Parameters']

    # Component 1: Sheet protection is enabled on 'Parameters' (0.3 points)
    # Initial state: protection.sheet == False; Golden state: protection.sheet == True
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 - Sheet protection is enabled (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 - Sheet protection is NOT enabled (protection.sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Password matches 'readonly123' (0.2 points)
    # Compute expected hash using openpyxl's built-in hashing
    # Initial state: password == None; Golden state: password hash == 'A11A'
    try:
        expected_sp = SheetProtection(password='readonly123', sheet=True)
        expected_hash = expected_sp.password  # 'A11A'
        actual_hash = ws.protection.password

        if actual_hash and actual_hash == expected_hash:
            print(f"PASS: Component 2 - Password hash matches 'readonly123' (hash={actual_hash}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 - Password hash mismatch. Expected '{expected_hash}', got '{actual_hash}'")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: All cells B2:B10 are unlocked/unprotected (0.3 points)
    # Initial state: all B2:B10 are locked=True; Golden state: all B2:B10 are locked=False
    try:
        unlocked_count = 0
        total_cells = 9  # B2 through B10
        for row in range(2, 11):
            cell = ws.cell(row=row, column=2)
            if not cell.protection.locked:
                unlocked_count += 1
            else:
                print(f"  INFO: B{row} is still locked (should be unlocked)")

        if unlocked_count == total_cells:
            print(f"PASS: Component 3 - All 9 cells B2:B10 are unlocked (0.3 pts)")
            total_score += 0.3
        elif unlocked_count > 0:
            # Partial credit: proportional to how many cells are correctly unlocked
            partial = round(0.3 * (unlocked_count / total_cells), 2)
            print(f"PARTIAL: Component 3 - {unlocked_count}/{total_cells} cells unlocked ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No cells in B2:B10 are unlocked ({unlocked_count}/{total_cells})")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Other cells remain locked (0.2 points)
    # Check a representative sample of non-B2:B10 cells to confirm they are locked
    # Initial state: all cells locked=True (but sheet protection off, so irrelevant)
    # Golden state: non-B2:B10 cells locked=True AND sheet protection on
    # This component only awards points if BOTH sheet protection is on AND other cells are locked
    try:
        sample_coords = ['A1', 'A2', 'A5', 'A10', 'B1', 'B11', 'C2', 'C5', 'D2', 'D5']
        locked_count = 0
        for coord in sample_coords:
            cell = ws[coord]
            if cell.protection.locked:
                locked_count += 1
            else:
                print(f"  INFO: {coord} is unlocked (should be locked)")

        # Only award points if sheet protection is active (prevents scoring initial state)
        if ws.protection.sheet and locked_count == len(sample_coords):
            print(f"PASS: Component 4 - All {len(sample_coords)} sampled non-B2:B10 cells are locked with protection active (0.2 pts)")
            total_score += 0.2
        elif ws.protection.sheet and locked_count > 0:
            partial = round(0.2 * (locked_count / len(sample_coords)), 2)
            print(f"PARTIAL: Component 4 - {locked_count}/{len(sample_coords)} sampled cells locked ({partial} pts)")
            total_score += partial
        else:
            if not ws.protection.sheet:
                print(f"FAIL: Component 4 - Sheet protection not active, locked status irrelevant")
            else:
                print(f"FAIL: Component 4 - No sampled non-B2:B10 cells are locked")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
