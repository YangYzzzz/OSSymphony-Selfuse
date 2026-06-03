"""
Initial Setup: Enable sheet-level protection on the Parameters sheet
Task ID: calc_gg3_034
Domain: libreoffice_calc

Creates a spreadsheet with a Parameters sheet containing labels in A2:A10,
editable input values in B2:B10, and formulas in other cells. No protection
is applied - that is the agent's task.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Parameters Sheet ---
    ws = wb.active
    ws.title = 'Parameters'

    # Header row styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    # Headers
    headers = {
        'A1': 'Parameter',
        'B1': 'Value',
        'C1': 'Unit',
        'D1': 'Description',
        'E1': 'Adjusted Value',
    }
    for coord, text in headers.items():
        cell = ws[coord]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Parameter labels (A2:A10) - fixed constants
    labels = [
        'Annual Growth Rate',
        'Discount Factor',
        'Tax Rate',
        'Inflation Adjustment',
        'Base Revenue',
        'Operating Margin',
        'Capital Expenditure',
        'Depreciation Rate',
        'Working Capital Ratio',
    ]

    # Input values (B2:B10) - editable by users
    values = [
        0.085,    # 8.5% growth
        0.92,     # discount factor
        0.21,     # 21% tax
        0.032,    # 3.2% inflation
        1250000,  # $1.25M base revenue
        0.185,    # 18.5% margin
        340000,   # $340K capex
        0.125,    # 12.5% depreciation
        0.15,     # 15% working capital
    ]

    # Units (C2:C10) - descriptive
    units = [
        '%', 'ratio', '%', '%', 'USD', '%', 'USD', '%', '%'
    ]

    # Descriptions (D2:D10) - explanatory text
    descriptions = [
        'Year-over-year revenue growth percentage',
        'Net present value discount factor',
        'Effective corporate tax rate',
        'Annual CPI-based inflation adjustment',
        'Baseline annual revenue for projections',
        'EBIT as percentage of revenue',
        'Annual capital expenditure budget',
        'Straight-line depreciation rate on fixed assets',
        'Net working capital as percentage of revenue',
    ]

    # Number formats for B column
    num_formats = [
        '0.00%', '0.00', '0.00%', '0.00%', '#,##0', '0.00%', '#,##0', '0.00%', '0.00%'
    ]

    label_font = Font(name='Calibri', size=11)
    value_font = Font(name='Calibri', size=11)
    value_fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')

    for i in range(9):
        row = i + 2

        # A column: labels
        cell_a = ws.cell(row=row, column=1, value=labels[i])
        cell_a.font = label_font
        cell_a.border = thin_border

        # B column: values (editable inputs)
        cell_b = ws.cell(row=row, column=2, value=values[i])
        cell_b.font = value_font
        cell_b.fill = value_fill
        cell_b.border = thin_border
        cell_b.number_format = num_formats[i]

        # C column: units
        cell_c = ws.cell(row=row, column=3, value=units[i])
        cell_c.font = label_font
        cell_c.border = thin_border
        cell_c.alignment = Alignment(horizontal='center')

        # D column: descriptions
        cell_d = ws.cell(row=row, column=4, value=descriptions[i])
        cell_d.font = label_font
        cell_d.border = thin_border

        # E column: formulas referencing B column (should NOT be overwritten)
        if num_formats[i] == '0.00%':
            formula = f'=B{row}*100'
            ws.cell(row=row, column=5, value=formula)
        elif num_formats[i] == '0.00':
            formula = f'=ROUND(B{row},4)'
            ws.cell(row=row, column=5, value=formula)
        else:
            formula = f'=B{row}*1.1'
            ws.cell(row=row, column=5, value=formula)
        cell_e = ws.cell(row=row, column=5)
        cell_e.font = label_font
        cell_e.border = thin_border

    # Add summary formulas at row 12
    ws['A12'] = 'Summary Statistics'
    ws['A12'].font = Font(name='Calibri', size=11, bold=True)
    ws['A13'] = 'Projected Revenue (Year 1)'
    ws['B13'] = '=B6*(1+B2)'
    ws['B13'].number_format = '#,##0.00'
    ws['A14'] = 'After-Tax Revenue'
    ws['B14'] = '=B13*(1-B4)'
    ws['B14'].number_format = '#,##0.00'
    ws['A15'] = 'Net after Inflation'
    ws['B15'] = '=B14*(1-B5)'
    ws['B15'].number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 18

    # --- Calculations Sheet (contains formulas referencing Parameters) ---
    ws2 = wb.create_sheet('Calculations')
    ws2['A1'] = 'Year'
    ws2['B1'] = 'Projected Revenue'
    ws2['C1'] = 'EBIT'
    ws2['D1'] = 'After Tax'
    for col_letter in ['A', 'B', 'C', 'D']:
        ws2[f'{col_letter}1'].font = header_font
        ws2[f'{col_letter}1'].fill = header_fill
        ws2[f'{col_letter}1'].alignment = header_align

    for yr in range(1, 6):
        row = yr + 1
        ws2.cell(row=row, column=1, value=yr)
        ws2.cell(row=row, column=2, value=f"=Parameters!B6*(1+Parameters!B2)^{yr}")
        ws2[f'B{row}'].number_format = '#,##0.00'
        ws2.cell(row=row, column=3, value=f"=B{row}*Parameters!B7")
        ws2[f'C{row}'].number_format = '#,##0.00'
        ws2.cell(row=row, column=4, value=f"=C{row}*(1-Parameters!B4)")
        ws2[f'D{row}'].number_format = '#,##0.00'

    ws2.column_dimensions['A'].width = 10
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 18
    ws2.column_dimensions['D'].width = 18

    # --- Summary Sheet ---
    ws3 = wb.create_sheet('Summary')
    ws3['A1'] = 'Model Summary'
    ws3['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws3['A3'] = 'This financial model uses parameters defined in the Parameters sheet.'
    ws3['A4'] = 'Users should modify values in B2:B10 of the Parameters sheet to run scenarios.'
    ws3['A5'] = 'All calculations update automatically based on parameter changes.'
    ws3['A7'] = 'Note: Do not modify formula cells directly.'

    # NO PROTECTION APPLIED - that is the agent's task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
