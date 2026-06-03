"""
Reward Script: Apply IF(ISBLANK()) formula in column D for conditional discount pricing
Task ID: calc_fma_if_isblank_055
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): All 12 cells D2:D13 contain non-empty formulas (any IF/ISBLANK formula)
  - Component 2 (0.3): Formulas follow the exact pattern =IF(ISBLANK(Cn),Bn,Bn*(1-Cn)) for each row
  - Component 3 (0.2): Formulas in D2 and D3 correctly reference their respective B and C columns (compound check that FAILS on initial)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_if_isblank_055'


def normalize_formula(formula):
    """Normalize formula string for comparison: uppercase, no spaces."""
    if formula is None:
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check the Invoice sheet exists
    if 'Invoice' not in wb.sheetnames:
        print("FAIL: Sheet 'Invoice' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Invoice']

    # Component 1: All 12 cells D2:D13 contain IF(ISBLANK(...)) formulas (0.5 points)
    # These cells should go from None (initial) to containing formulas (golden).
    try:
        formula_cells_found = 0
        formula_cells_with_isblank = 0
        missing_rows = []
        non_isblank_rows = []

        for row in range(2, 14):  # rows 2-13 inclusive
            d_val = ws.cell(row=row, column=4).value
            if d_val is not None and isinstance(d_val, str) and d_val.startswith('='):
                formula_cells_found += 1
                normalized = normalize_formula(d_val)
                if 'ISBLANK' in normalized and 'IF' in normalized:
                    formula_cells_with_isblank += 1
                else:
                    non_isblank_rows.append(row)
            else:
                missing_rows.append(row)

        if formula_cells_found == 12 and formula_cells_with_isblank == 12:
            print(f"PASS: Component 1 — All 12 cells D2:D13 contain IF(ISBLANK(...)) formulas (0.5 pts)")
            total_score += 0.5
        elif formula_cells_found == 12:
            # All 12 have formulas but some don't use ISBLANK pattern
            print(f"PARTIAL: Component 1 — All 12 cells have formulas but {len(non_isblank_rows)} don't use ISBLANK: rows {non_isblank_rows}")
            # Award half credit for having all formulas present even if not ISBLANK pattern
            print(f"  (0.25 pts partial)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Only {formula_cells_found}/12 cells have formulas. Missing in rows: {missing_rows}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formulas follow the exact expected pattern for each row (0.3 points)
    # Pattern: =IF(ISBLANK(Cn),Bn,Bn*(1-Cn)) where n is the row number
    try:
        exact_match_count = 0
        wrong_formulas = []

        for row in range(2, 14):
            d_val = ws.cell(row=row, column=4).value
            if d_val is not None and isinstance(d_val, str):
                normalized = normalize_formula(d_val)
                # Expected: =IF(ISBLANK(C{row}),B{row},B{row}*(1-C{row}))
                expected_pattern = normalize_formula(f'=IF(ISBLANK(C{row}),B{row},B{row}*(1-C{row}))')
                if normalized == expected_pattern:
                    exact_match_count += 1
                else:
                    # Also check alternative correct formula variants
                    # Some agents may write: =IF(C2="",B2,B2*(1-C2)) or =IF(C2<>"",B2*(1-C2),B2)
                    # Check if the formula logically handles the blank-discount pattern
                    # by checking for presence of correct cell references
                    alt_isblank = normalize_formula(f'=IF(ISBLANK(C{row}),B{row},B{row}*(1-C{row}))')
                    alt_empty_str = normalize_formula(f'=IF(C{row}="",B{row},B{row}*(1-C{row}))')
                    if normalized in [alt_isblank, alt_empty_str]:
                        exact_match_count += 1
                    else:
                        wrong_formulas.append((row, d_val))

        if exact_match_count == 12:
            print(f"PASS: Component 2 — All 12 formulas match the exact expected pattern =IF(ISBLANK(Cn),Bn,Bn*(1-Cn)) (0.3 pts)")
            total_score += 0.3
        elif exact_match_count >= 6:
            print(f"PARTIAL: Component 2 — {exact_match_count}/12 formulas match exact pattern. Wrong: {wrong_formulas[:3]}")
            print(f"  (0.15 pts partial)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Only {exact_match_count}/12 formulas match. Wrong formulas: {wrong_formulas[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas correctly handle BOTH discount and non-discount rows (0.2 points)
    # This checks the LOGIC is correct — verifies specific rows with and without discount.
    # A discount row: D2 formula references C2 (non-blank). A non-discount row: D3 references C3 (blank).
    # This FAILS on the initial file (D2/D3 are None) and PASSES on golden (formulas present and typed correctly).
    # Specifically: D2 must use the discount (references both B2 and C2), D3 must handle blank (references B3, C3 via ISBLANK).
    try:
        # Check D2: row with discount (C2=0.10) — formula must reference C2 and B2 for discounted price
        d2_val = ws.cell(row=2, column=4).value
        # Check D3: row WITHOUT discount (C3=None) — formula must reference C3 and B3 for full price
        d3_val = ws.cell(row=3, column=4).value

        d2_is_formula = (d2_val is not None and isinstance(d2_val, str) and d2_val.startswith('='))
        d3_is_formula = (d3_val is not None and isinstance(d3_val, str) and d3_val.startswith('='))

        if d2_is_formula and d3_is_formula:
            # Both representative cells have formulas — now check they reference the right cells
            d2_norm = normalize_formula(d2_val)
            d3_norm = normalize_formula(d3_val)

            # D2 should reference both B2 and C2 (for discounted calculation)
            d2_has_b2 = 'B2' in d2_norm
            d2_has_c2 = 'C2' in d2_norm
            # D3 should reference both B3 and C3 (C3 is checked via ISBLANK)
            d3_has_b3 = 'B3' in d3_norm
            d3_has_c3 = 'C3' in d3_norm

            if d2_has_b2 and d2_has_c2 and d3_has_b3 and d3_has_c3:
                print(f"PASS: Component 3 — Formulas correctly reference price and discount for both discount/non-discount rows (0.2 pts)")
                print(f"  D2: {d2_val}, D3: {d3_val}")
                total_score += 0.2
            else:
                missing = []
                if not d2_has_b2: missing.append("D2 missing B2 ref")
                if not d2_has_c2: missing.append("D2 missing C2 ref")
                if not d3_has_b3: missing.append("D3 missing B3 ref")
                if not d3_has_c3: missing.append("D3 missing C3 ref")
                print(f"FAIL: Component 3 — Formula cell reference issues: {missing}")
        else:
            print(f"FAIL: Component 3 — D2 or D3 is not a formula. D2={repr(d2_val)}, D3={repr(d3_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
