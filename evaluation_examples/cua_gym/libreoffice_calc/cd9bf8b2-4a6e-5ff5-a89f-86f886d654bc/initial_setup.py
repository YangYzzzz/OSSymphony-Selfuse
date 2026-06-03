"""
Initial Setup: Invoice table with discount percentage column - Final Price column empty
Task ID: calc_fma_if_isblank_055
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_if_isblank_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Invoice ---
    ws = wb.active
    ws.title = 'Invoice'

    # Header row styling
    header_font = Font(name='Calibri', bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers: A1=Item, B1=Unit Price, C1=Discount, D1=Final Price
    headers = ['Item', 'Unit Price', 'Discount', 'Final Price']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', bold=True, size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Data: item names, unit prices, discount percentages (some blank)
    # Column C discounts: 0.10, blank, 0.15, 0.20, blank, 0.05, blank, 0.10, blank, 0.25, blank, 0.30
    items = [
        'Wireless Mouse',
        'USB-C Hub',
        'HDMI Cable',
        'Monitor Stand',
        'Keyboard Cover',
        'Sticky Notes Pack',
        'Laptop Sleeve',
        'Screen Cleaner',
        'Cable Organizer',
        'Ergonomic Chair',
        'Pen Set',
        'Webcam',
    ]
    prices = [25.00, 89.99, 12.50, 199.00, 45.00, 8.75, 150.00, 32.99, 18.50, 275.00, 5.99, 120.00]
    discounts = [0.10, None, 0.15, 0.20, None, 0.05, None, 0.10, None, 0.25, None, 0.30]

    data_align = Alignment(horizontal='left', vertical='center')
    price_align = Alignment(horizontal='right', vertical='center')

    for i, (item, price, disc) in enumerate(zip(items, prices, discounts)):
        row = i + 2
        # Column A: Item name
        cell_a = ws.cell(row=row, column=1, value=item)
        cell_a.alignment = data_align
        cell_a.border = border

        # Column B: Unit price
        cell_b = ws.cell(row=row, column=2, value=price)
        cell_b.number_format = '#,##0.00'
        cell_b.alignment = price_align
        cell_b.border = border

        # Column C: Discount (leave blank if None)
        cell_c = ws.cell(row=row, column=3)
        if disc is not None:
            cell_c.value = disc
            cell_c.number_format = '0%'
        cell_c.alignment = price_align
        cell_c.border = border

        # Column D: Final Price — intentionally empty (task requires adding formula here)
        cell_d = ws.cell(row=row, column=4)
        cell_d.border = border
        cell_d.alignment = price_align
        cell_d.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14

    # Row 1 height
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
