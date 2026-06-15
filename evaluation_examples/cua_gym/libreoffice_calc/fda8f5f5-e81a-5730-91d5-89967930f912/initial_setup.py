"""
Initial Setup: Q2 2026 Schedule spreadsheet with date column but no validation
Task ID: calc_gcv_081
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Q2_Schedule"

    # --- Headers ---
    headers = ["Activity ID", "Activity", "Responsible", "Scheduled Date"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- 29 activities for Q2 planning ---
    activities = [
        ("ACT-001", "Q2 Budget Review", "Sarah Chen"),
        ("ACT-002", "Marketing Campaign Launch", "Marcus Johnson"),
        ("ACT-003", "Product Roadmap Presentation", "Emily Rodriguez"),
        ("ACT-004", "Customer Feedback Analysis", "David Kim"),
        ("ACT-005", "Supply Chain Optimization Meeting", "Lisa Wang"),
        ("ACT-006", "New Hire Onboarding Session", "James Thompson"),
        ("ACT-007", "Quarterly Sales Target Setting", "Priya Patel"),
        ("ACT-008", "IT Infrastructure Upgrade Planning", "Robert Garcia"),
        ("ACT-009", "Employee Engagement Survey", "Aisha Okonkwo"),
        ("ACT-010", "Vendor Contract Renewal", "Michael Brown"),
        ("ACT-011", "Product Quality Audit", "Yuki Tanaka"),
        ("ACT-012", "Team Building Workshop", "Natalie Foster"),
        ("ACT-013", "Risk Assessment Review", "Carlos Mendez"),
        ("ACT-014", "Compliance Training", "Hannah Lee"),
        ("ACT-015", "R&D Sprint Planning", "Oliver Schmidt"),
        ("ACT-016", "Client Portfolio Review", "Sophia Martinez"),
        ("ACT-017", "Facility Maintenance Scheduling", "Ahmed Hassan"),
        ("ACT-018", "Data Migration Project Kickoff", "Rachel Cooper"),
        ("ACT-019", "Social Media Strategy Update", "Tyler Nguyen"),
        ("ACT-020", "Financial Reconciliation", "Victoria Adams"),
        ("ACT-021", "Safety Compliance Inspection", "Daniel Wright"),
        ("ACT-022", "Product Launch Rehearsal", "Megan Clark"),
        ("ACT-023", "Inventory Stocktake", "Kevin Zhao"),
        ("ACT-024", "Partnership Agreement Signing", "Isabelle Dubois"),
        ("ACT-025", "Customer Success Review", "Nathan Brooks"),
        ("ACT-026", "Training Materials Update", "Grace Osei"),
        ("ACT-027", "End-of-Q2 Reporting Prep", "William Turner"),
        ("ACT-028", "Sustainability Initiative Launch", "Elena Volkov"),
        ("ACT-029", "Performance Review Calibration", "Christopher Lee"),
    ]

    data_font = Font(name="Calibri", size=11)
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r, (act_id, activity, responsible) in enumerate(activities, 2):
        for c, val in enumerate([act_id, activity, responsible], 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = data_border
            if c == 1:
                cell.alignment = Alignment(horizontal="center")

        # Column D: empty but formatted as date
        d_cell = ws.cell(row=r, column=4)
        d_cell.number_format = 'yyyy-mm-dd'
        d_cell.border = data_border
        d_cell.alignment = Alignment(horizontal="center")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
