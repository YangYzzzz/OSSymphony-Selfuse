"""
Initial Setup: Batch document conversion quality-check pipeline
Task ID: osworld_multi_apps_doc_batch_convert_012
Domain: multi_apps (LibreOffice Writer + Calc + OS)

Creates:
  - /home/user/docs_input/  with 12 .odt files (realistic business documents)
  - /home/user/Desktop/expected_manifest.ods  with Filename + Expected_Page_Count
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_doc_batch_convert_012'
DOCS_INPUT = f'{WORKDIR}/docs_input'
DESKTOP = f'{WORKDIR}/Desktop'
MANIFEST_PATH = f'{DESKTOP}/expected_manifest.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_odt_document(filepath, title, content_pages):
    """
    Create an ODT document using LibreOffice headless via a temporary Python script.
    Each element in content_pages is a list of paragraphs for that page.
    """
    # Build the document content as plain text then use python-docx → convert via LO
    # Actually we create .odt directly using python-docx → .docx then save as odt via LO headless
    # Simpler: write a script that uses odfpy to create odt
    from odf.opendocument import OpenDocumentText
    from odf.style import Style, TextProperties, ParagraphProperties, PageLayout, MasterPage
    from odf.text import P, H, Span
    from odf.draw import Frame, TextBox

    doc = OpenDocumentText()

    # Create heading style
    h1_style = Style(name="Heading1", family="paragraph")
    h1_style.addElement(TextProperties(fontsize="18pt", fontweight="bold"))
    doc.styles.addElement(h1_style)

    # Create body style
    body_style = Style(name="BodyText", family="paragraph")
    body_style.addElement(TextProperties(fontsize="12pt"))
    doc.styles.addElement(body_style)

    # Add title heading
    heading = H(outlinelevel=1, stylename=h1_style)
    heading.addText(title)
    doc.text.addElement(heading)

    for page_idx, page_paragraphs in enumerate(content_pages):
        for para_text in page_paragraphs:
            p = P(stylename=body_style)
            p.addText(para_text)
            doc.text.addElement(p)

        # Add page break between pages (except last)
        if page_idx < len(content_pages) - 1:
            # Add multiple empty paragraphs to simulate a page break for pdfinfo
            # Use a paragraph with page-break style
            pb_style = Style(name=f"PageBreak{page_idx}", family="paragraph")
            pb_style.addElement(ParagraphProperties(breakbefore="page"))
            doc.automaticstyles.addElement(pb_style)
            pb_para = P(stylename=pb_style)
            pb_para.addText("")
            doc.text.addElement(pb_para)

    doc.save(filepath)


# Define 12 documents with realistic content and expected page counts
DOCUMENTS = [
    {
        "filename": "annual_report_2024.odt",
        "title": "Annual Report 2024 - Nexus Technologies",
        "expected_pages": 3,
        "pages": [
            [
                "Executive Summary",
                "Nexus Technologies delivered strong financial performance in fiscal year 2024, achieving record revenue of $142.7 million, a 23% increase over the prior year. Operating income reached $28.4 million, reflecting improved operational efficiency and disciplined cost management.",
                "Net income attributable to shareholders was $21.3 million, or $1.87 per diluted share, compared to $16.8 million, or $1.48 per diluted share, in 2023. The board of directors approved a quarterly dividend of $0.15 per share.",
            ],
            [
                "Business Segments Overview",
                "Cloud Services Division: Revenue grew 31% year-over-year to $89.4 million, driven by increased enterprise adoption of our NexCloud platform. New contract signings totaled 847 enterprise customers, with an average contract value of $105,600.",
                "Professional Services Division: Revenue increased 11% to $53.3 million. Utilization rates averaged 84% across our global delivery centers in Austin, Dublin, Singapore, and Nairobi.",
                "Research and Development: We invested $18.6 million in R&D, representing 13% of revenue. Key initiatives include next-generation AI orchestration, edge computing modules, and quantum-resistant encryption protocols.",
            ],
            [
                "Financial Highlights and Outlook",
                "Balance Sheet: Total assets reached $318.5 million at year-end, including $67.2 million in cash and short-term investments. Long-term debt stands at $45.0 million following the refinancing of our revolving credit facility.",
                "2025 Guidance: Management expects revenue in the range of $163 to $168 million, representing growth of 14-18%. Operating margin is projected to expand by 80-120 basis points.",
                "Risk Factors: Principal risks include macroeconomic uncertainty, cybersecurity threats, talent retention, and competitive pressure from hyperscale cloud providers.",
            ],
        ],
    },
    {
        "filename": "employee_handbook_v3.odt",
        "title": "Employee Handbook Version 3.0 - Meridian Financial Group",
        "expected_pages": 3,
        "pages": [
            [
                "Welcome and Mission Statement",
                "Welcome to Meridian Financial Group. Our mission is to empower individuals and businesses to achieve their financial aspirations through transparent, ethical, and innovative financial services.",
                "This handbook outlines the policies, procedures, and expectations that guide our work environment. All employees are expected to familiarise themselves with these guidelines upon joining.",
            ],
            [
                "Code of Conduct and Ethics",
                "Meridian Financial Group holds all employees to the highest standards of professional conduct. Conflicts of interest must be disclosed promptly to your line manager and the Compliance team.",
                "Insider trading, market manipulation, and any form of financial fraud are strictly prohibited and will result in immediate termination and referral to regulatory authorities.",
                "Personal account dealings in securities must be pre-approved by the Compliance Department. Employees must wait 24 hours after client trades before executing personal trades in the same securities.",
            ],
            [
                "Benefits and Leave Policy",
                "Health Insurance: Meridian provides comprehensive medical, dental, and vision coverage for all full-time employees and eligible dependants. Premiums are 80% employer-funded.",
                "Paid Time Off: Employees accrue 15 days of PTO per year in years 1-3, increasing to 20 days in years 4-7, and 25 days thereafter. Unused PTO up to 5 days may be carried forward.",
                "Parental Leave: Primary caregivers receive 16 weeks of fully paid parental leave. Secondary caregivers receive 4 weeks of fully paid leave.",
            ],
        ],
    },
    {
        "filename": "project_charter_alpha.odt",
        "title": "Project Charter - Project Alpha - Streamline Operations Initiative",
        "expected_pages": 2,
        "pages": [
            [
                "Project Overview and Objectives",
                "Project Alpha is a strategic initiative to streamline operational workflows across the Supply Chain and Finance departments of Hartwell Manufacturing Inc. The project is sponsored by the Chief Operating Officer and overseen by the Project Management Office.",
                "Primary Objective: Reduce order-to-cash cycle time from an average of 18 days to 11 days by Q3 2025.",
                "Secondary Objectives: Automate manual data entry for 75% of purchase orders; implement real-time inventory dashboards; reduce month-end close from 8 days to 5 days.",
            ],
            [
                "Scope, Timeline, and Budget",
                "In Scope: ERP system configuration (SAP S/4HANA), workflow automation (RPA bots), staff training, data migration from legacy Oracle system.",
                "Out of Scope: Customer-facing portal redesign; HR module upgrades; regional office rollouts outside North America.",
                "Timeline: Phase 1 (Discovery & Design) - January to March 2025; Phase 2 (Build & Test) - April to July 2025; Phase 3 (Deployment) - August to September 2025.",
                "Budget: Total approved budget is $2,350,000. Contingency reserve of 15% ($352,500) requires COO approval to release.",
            ],
        ],
    },
    {
        "filename": "training_materials_safety.odt",
        "title": "Workplace Safety Training Materials - Section 4: Chemical Handling",
        "expected_pages": 2,
        "pages": [
            [
                "Introduction to Chemical Safety",
                "This module covers the safe handling, storage, and disposal of hazardous chemicals used in our manufacturing facilities. Completion of this module is mandatory for all production floor staff and laboratory personnel.",
                "Relevant Regulations: OSHA Hazard Communication Standard (29 CFR 1910.1200), EPA Resource Conservation and Recovery Act (RCRA), and local fire and safety codes.",
                "Personal Protective Equipment (PPE): Chemical-resistant gloves (nitrile or neoprene), safety goggles with side shields, lab coat or chemical-resistant apron, and steel-toed footwear are required at all times in the chemical storage area.",
            ],
            [
                "Spill Response and Emergency Procedures",
                "Minor Spill (< 1 litre): Don appropriate PPE, contain spill with absorbent material, dispose of waste in labelled hazardous waste container, notify supervisor and complete incident report within 2 hours.",
                "Major Spill (> 1 litre) or Unknown Chemical: Evacuate the immediate area, activate the nearest fire alarm pull station, call Emergency Response at extension 911, do not attempt to clean up without HAZMAT team approval.",
                "Emergency Eye Wash Station: Located within 10 seconds travel distance of all chemical work areas. Flush affected eyes for a minimum of 15 continuous minutes. Seek immediate medical attention after flushing.",
            ],
        ],
    },
    {
        "filename": "meeting_minutes_q1_review.odt",
        "title": "Meeting Minutes - Q1 2025 Business Review - Oakdale Retail Group",
        "expected_pages": 1,
        "pages": [
            [
                "Attendees and Agenda",
                "Date: March 15, 2025 | Time: 09:00 - 11:30 | Location: Boardroom 3A, Oakdale HQ",
                "Attendees: Margaret Thornton (CEO), David Osei (CFO), Rachel Kim (CMO), James Calloway (VP Operations), Sofia Mendez (Head of E-Commerce), Thomas Brennan (Head of Store Operations).",
                "Agenda: (1) Q1 financial results review; (2) E-commerce performance update; (3) Store operations highlights; (4) Marketing campaign debrief; (5) Q2 priorities; (6) AOB.",
                "Q1 Financial Results: Total revenue reached $48.3 million against a budget of $45.0 million (+7.3%). Gross margin improved to 41.2% from 39.8% in Q1 2024. EBITDA was $7.9 million (16.4% margin).",
                "Action Items: D. Osei to circulate updated cash flow forecast by March 22. R. Kim to present Q2 campaign spend reallocation proposal by March 29. J. Calloway to investigate logistics costs spike in the Southwest region.",
            ],
        ],
    },
    {
        "filename": "product_specifications_v2.odt",
        "title": "Product Specifications v2.0 - ThermoCore Pro Series",
        "expected_pages": 2,
        "pages": [
            [
                "Product Overview",
                "The ThermoCore Pro Series is a line of industrial-grade thermal management solutions designed for high-power electronics in aerospace, defence, and telecommunications applications.",
                "Model Range: TCP-250 (250W), TCP-500 (500W), TCP-1000 (1000W), TCP-2000 (2000W).",
                "Operating Temperature Range: -55°C to +125°C ambient. Storage Temperature: -65°C to +150°C.",
                "Thermal Resistance: TCP-250: 0.18°C/W; TCP-500: 0.09°C/W; TCP-1000: 0.045°C/W; TCP-2000: 0.022°C/W.",
            ],
            [
                "Mechanical and Electrical Specifications",
                "Dimensions (L x W x H): TCP-250: 95mm x 60mm x 22mm; TCP-500: 140mm x 75mm x 28mm; TCP-1000: 195mm x 100mm x 35mm; TCP-2000: 260mm x 130mm x 45mm.",
                "Weight: TCP-250: 285g; TCP-500: 620g; TCP-1000: 1,350g; TCP-2000: 3,100g.",
                "Input Voltage: 24VDC nominal (18-32VDC range). Efficiency: >93% at full load for all models.",
                "Certifications: MIL-STD-810H (Environmental), MIL-STD-461G (EMC), UL 60950-1, CE Mark (LVD and EMC Directives).",
            ],
        ],
    },
    {
        "filename": "research_summary_biotech.odt",
        "title": "Research Summary - Novel Enzyme Inhibitors for Inflammatory Disease",
        "expected_pages": 2,
        "pages": [
            [
                "Abstract and Background",
                "This document summarises Phase II clinical trial results for compound NX-4471, a selective JAK1/TYK2 inhibitor developed for moderate-to-severe rheumatoid arthritis.",
                "Background: Current JAK inhibitor therapies demonstrate efficacy but carry class-related risks including elevated lipid levels, thromboembolism, and increased susceptibility to infections. NX-4471 was designed to achieve greater selectivity to reduce off-target effects.",
                "Trial Design: Double-blind, placebo-controlled, multi-centre study across 28 sites in the United States, Germany, Japan, and Australia. 486 patients randomised 2:1 (active:placebo). Primary endpoint: ACR50 response at week 12.",
            ],
            [
                "Results and Discussion",
                "Efficacy: ACR50 response at week 12 was achieved by 61.4% of patients in the NX-4471 group versus 18.7% in the placebo group (p < 0.0001). ACR70 was achieved by 38.2% versus 8.3% respectively.",
                "Safety Profile: Treatment-emergent adverse events were reported in 52.1% of active and 48.3% of placebo patients. Serious adverse events: 4.7% active vs 5.1% placebo (non-significant difference).",
                "Lipid Effects: Mean LDL-C increase of 3.2 mg/dL in the active group versus 1.1 mg/dL in placebo; substantially lower than comparator JAK inhibitors (typically 8-12 mg/dL increase).",
                "Conclusion: NX-4471 demonstrates a statistically significant and clinically meaningful improvement in RA symptoms with a differentiated safety profile. Phase III initiation planned for Q4 2025.",
            ],
        ],
    },
    {
        "filename": "contract_template_services.odt",
        "title": "Master Services Agreement Template - Pinnacle Consulting LLC",
        "expected_pages": 2,
        "pages": [
            [
                "Parties and Recitals",
                "This Master Services Agreement (\"Agreement\") is entered into as of the Effective Date by and between Pinnacle Consulting LLC, a Delaware limited liability company (\"Service Provider\"), and the Client identified in the applicable Statement of Work (\"Client\").",
                "RECITALS: WHEREAS, Service Provider desires to provide certain consulting and professional services to Client; WHEREAS, Client desires to obtain such services from Service Provider; NOW, THEREFORE, in consideration of the mutual covenants and agreements set forth herein, the parties agree as follows.",
                "Definitions: \"Services\" means consulting, advisory, and project management services as described in each SOW. \"Deliverables\" means work product created specifically for Client under an SOW. \"Confidential Information\" means non-public information disclosed by either party.",
            ],
            [
                "Terms, Payment, and Liability",
                "Payment Terms: Client shall pay undisputed invoices within 30 days of receipt. Late payments accrue interest at 1.5% per month. Service Provider may suspend services after 15 days written notice of non-payment.",
                "Intellectual Property: All Deliverables created exclusively for Client and fully paid for shall be owned by Client upon payment in full. Pre-existing materials and tools remain the property of Service Provider; a non-exclusive licence is granted to Client for use of such materials as incorporated in Deliverables.",
                "Limitation of Liability: NEITHER PARTY SHALL BE LIABLE FOR INDIRECT, INCIDENTAL, SPECIAL, OR CONSEQUENTIAL DAMAGES. Service Provider's total liability shall not exceed the fees paid in the 3 months preceding the claim.",
            ],
        ],
    },
    {
        "filename": "budget_proposal_2025.odt",
        "title": "Budget Proposal FY2025 - Information Technology Department",
        "expected_pages": 2,
        "pages": [
            [
                "Executive Summary and Request",
                "The Information Technology Department submits this budget proposal for fiscal year 2025 totalling $4,875,000, representing a 12.7% increase over the FY2024 approved budget of $4,324,000.",
                "Key drivers of the increase include: cloud infrastructure expansion (+$310,000), cybersecurity enhancements (+$185,000), AI/ML tooling (+$145,000), and headcount for two additional senior engineers (+$220,000).",
                "Projected ROI: The proposed investments are expected to generate $1.2 million in productivity savings, $430,000 in avoided legacy maintenance costs, and reduce security incident response costs by an estimated $280,000.",
            ],
            [
                "Line Item Breakdown",
                "Personnel: Base salaries $2,180,000; benefits and taxes $545,000; contractor budget $320,000. Total Personnel: $3,045,000.",
                "Infrastructure: Cloud hosting (AWS/Azure) $890,000; on-premise hardware refresh $215,000; networking equipment $78,000. Total Infrastructure: $1,183,000.",
                "Software Licences: Enterprise applications $312,000; security tools $198,000; development tools $87,000; collaboration platforms $50,000. Total Software: $647,000.",
                "Training and Development: Technical certifications $45,000; conference attendance $28,000; e-learning platforms $22,000. Total T&D: $95,000. Contingency: $55,000.",
            ],
        ],
    },
    {
        "filename": "policy_document_remote_work.odt",
        "title": "Remote Work Policy - Crestwood Insurance Partners",
        "expected_pages": 1,
        "pages": [
            [
                "Policy Statement and Eligibility",
                "Crestwood Insurance Partners supports flexible work arrangements that maintain high performance standards while promoting employee wellbeing. This policy governs remote work for eligible employees.",
                "Eligibility: Employees who have completed their 90-day probationary period, whose roles are conducive to remote work, and who have received approval from their direct manager and HR Business Partner.",
                "Arrangement Types: (1) Hybrid - minimum 2 days in office per week (standard arrangement); (2) Primarily Remote - maximum 1 day in office per week (requires Director-level approval); (3) Fully Remote - no regular office presence (requires VP approval and documented business justification).",
                "Equipment and Security: Company laptops must be used for all work; personal devices may not access client data. VPN connection is mandatory when accessing company systems remotely. Home network must have WPA3 or WPA2 security enabled. Physical security of company equipment is the employee's responsibility.",
            ],
        ],
    },
    {
        "filename": "technical_guide_api_v4.odt",
        "title": "Technical Integration Guide - Helix Data API v4.0",
        "expected_pages": 3,
        "pages": [
            [
                "Introduction and Authentication",
                "The Helix Data API v4.0 provides programmatic access to real-time and historical market data, corporate fundamentals, and analytics. This guide covers authentication, endpoint reference, rate limiting, and best practices.",
                "Base URL: https://api.helixdata.io/v4. All requests must use HTTPS. HTTP requests are rejected with status 426 (Upgrade Required).",
                "Authentication: API key authentication via Bearer token in the Authorization header. Example: Authorization: Bearer hx4_live_a8f3k9p2m7n1q5r6s0t4w8.",
                "OAuth 2.0 is available for enterprise accounts. Redirect URI must be pre-registered. Token lifetime: access token 3600 seconds; refresh token 30 days.",
            ],
            [
                "Core Endpoints",
                "GET /quotes/{symbol} - Returns real-time quote. Parameters: symbol (required), fields (optional, comma-separated). Response includes bid, ask, last, volume, change, change_pct.",
                "GET /history/{symbol} - Returns OHLCV history. Parameters: symbol (required), from (YYYY-MM-DD), to (YYYY-MM-DD), interval (1m, 5m, 15m, 1h, 1d, 1w, 1mo).",
                "GET /fundamentals/{symbol} - Returns financial statement data. Parameters: symbol (required), type (income, balance, cashflow, ratios), period (annual, quarterly), limit (1-20).",
                "POST /screener - Screen securities by criteria. Body: JSON object with filter conditions, sort specification, and pagination (offset, limit).",
            ],
            [
                "Rate Limits and Error Handling",
                "Rate Limits: Basic plan: 100 requests/minute, 10,000/day. Professional plan: 500 requests/minute, 100,000/day. Enterprise: custom. Limit headers: X-RateLimit-Limit, X-RateLimit-Remaining, X-RateLimit-Reset.",
                "Error Codes: 400 Bad Request (invalid parameters), 401 Unauthorized (invalid/expired API key), 403 Forbidden (plan limit exceeded), 404 Not Found (symbol not found), 429 Too Many Requests (rate limit exceeded), 500 Internal Server Error (contact support).",
                "Best Practices: Implement exponential back-off for 429 errors starting at 1 second. Cache responses where possible (quote data: 1 second minimum; fundamentals: 24 hours). Use websocket endpoint /ws/stream for real-time data to avoid polling.",
            ],
        ],
    },
    {
        "filename": "compliance_report_gdpr.odt",
        "title": "GDPR Compliance Assessment Report 2024 - Veridian Data Systems",
        "expected_pages": 3,
        "pages": [
            [
                "Executive Summary",
                "This report presents the findings of the annual GDPR compliance assessment conducted by the Data Protection team in collaboration with external auditors Grant & Whitmore LLP.",
                "Overall Compliance Status: SUBSTANTIALLY COMPLIANT. 47 of 52 assessed controls rated as Effective. 5 controls rated as Partially Effective with remediation plans in place. 0 controls rated as Ineffective.",
                "Key Findings: (1) Data Subject Rights processes are well-established with 98.4% of requests fulfilled within the statutory 30-day period. (2) Privacy by Design principles are embedded in the SDLC. (3) Data Protection Impact Assessments completed for all new high-risk processing activities.",
            ],
            [
                "Data Inventory and Legal Bases",
                "Data Categories Processed: Personal identifiers (name, email, address), financial data (payment card, bank account), behavioural data (usage logs, preferences), and special category data (health information for 3 specific product lines).",
                "Legal Bases: Contractual necessity (Article 6(1)(b)): 62% of processing activities. Legitimate interests (Article 6(1)(f)): 28% of processing activities. Consent (Article 6(1)(a)): 10% of processing activities, primarily marketing communications.",
                "Third-Party Data Transfers: 14 active data processing agreements with sub-processors. 3 sub-processors located outside the EEA; all covered by Standard Contractual Clauses (SCCs) updated post-Schrems II.",
            ],
            [
                "Remediation Actions and Roadmap",
                "Partially Effective Controls Requiring Remediation: (1) Data Retention - 3 legacy systems lack automated deletion; target completion Q2 2025. (2) Consent Management Platform - cookie consent not fully IAB TCF 2.2 compliant; vendor patch expected Q1 2025.",
                "Upcoming Regulatory Changes: EU AI Act obligations will require review of automated decision-making processes by Q4 2025. Data governance team to assess applicability and required controls.",
                "Training Completion: 94.7% of staff completed mandatory data protection training in 2024. Non-compliant staff (5.3%) are in the process of completing overdue training; escalation to line managers initiated.",
            ],
        ],
    },
]


def create_manifest():
    """Create expected_manifest.ods using openpyxl (save as xlsx, then we need odfpy for ods)."""
    try:
        import pyods
    except ImportError:
        pass

    # Use openpyxl to create the file content, then save as .ods using ezodf or odfpy
    # Actually, let's use pyexcel-ods or just write it with odfpy directly
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties, TableCellProperties

        doc = OpenDocumentSpreadsheet()

        # Create header style
        header_style = Style(name="HeaderStyle", family="table-cell")
        header_style.addElement(TextProperties(fontweight="bold"))
        doc.styles.addElement(header_style)

        table = Table(name="Manifest")

        # Header row
        header_row = TableRow()
        for header_text in ["Filename", "Expected_Page_Count"]:
            cell = TableCell(stylename=header_style)
            cell.addElement(P(text=header_text))
            header_row.addElement(cell)
        table.addElement(header_row)

        # Data rows
        for doc_info in DOCUMENTS:
            row = TableRow()
            # Filename cell
            fname_cell = TableCell()
            fname_cell.addElement(P(text=doc_info["filename"]))
            row.addElement(fname_cell)
            # Page count cell
            pcount_cell = TableCell(valuetype="float", value=str(doc_info["expected_pages"]))
            pcount_cell.addElement(P(text=str(doc_info["expected_pages"])))
            row.addElement(pcount_cell)
            table.addElement(row)

        doc.spreadsheet.addElement(table)
        doc.save(MANIFEST_PATH)
        print(f"Manifest created: {MANIFEST_PATH}")

    except Exception as e:
        print(f"odfpy ODS creation failed: {e}. Falling back to openpyxl xlsx.")
        # Fallback: save as xlsx with .ods extension (LO can read it)
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Manifest"
        ws.cell(row=1, column=1, value="Filename")
        ws.cell(row=1, column=2, value="Expected_Page_Count")
        for i, doc_info in enumerate(DOCUMENTS, 2):
            ws.cell(row=i, column=1, value=doc_info["filename"])
            ws.cell(row=i, column=2, value=doc_info["expected_pages"])
        wb.save(MANIFEST_PATH)
        print(f"Manifest created (xlsx-as-ods fallback): {MANIFEST_PATH}")


def create_initial():
    # Create docs_input directory
    os.makedirs(DOCS_INPUT, exist_ok=True)
    print(f"Created directory: {DOCS_INPUT}")

    # Ensure Desktop exists
    os.makedirs(DESKTOP, exist_ok=True)

    # Create 12 .odt files
    for doc_info in DOCUMENTS:
        filepath = os.path.join(DOCS_INPUT, doc_info["filename"])
        try:
            create_odt_document(filepath, doc_info["title"], doc_info["pages"])
            print(f"Created: {filepath}")
        except Exception as e:
            print(f"ERROR creating {filepath}: {e}")
            raise

    # Create the manifest
    create_manifest()

    # Verify files
    created = [f for f in os.listdir(DOCS_INPUT) if f.endswith('.odt')]
    print(f"\nCreated {len(created)} ODT files in {DOCS_INPUT}:")
    for f in sorted(created):
        size = os.path.getsize(os.path.join(DOCS_INPUT, f))
        print(f"  {f} ({size} bytes)")

    print(f"\nManifest at: {MANIFEST_PATH}")
    if os.path.exists(MANIFEST_PATH):
        size = os.path.getsize(MANIFEST_PATH)
        print(f"  Size: {size} bytes")

    # GUI-ready startup: open the manifest in LibreOffice Calc
    # and a file manager showing docs_input/
    launch_gui(f'libreoffice --calc "{MANIFEST_PATH}"', delay_sec=2.0)
    launch_gui(f'nautilus "{DOCS_INPUT}"', delay_sec=1.0)
    print("GUI_READY: launched LibreOffice Calc with manifest and Nautilus with docs_input/")


create_initial()
