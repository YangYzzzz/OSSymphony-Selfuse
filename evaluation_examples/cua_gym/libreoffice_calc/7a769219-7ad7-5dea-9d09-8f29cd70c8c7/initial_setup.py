"""
Initial Setup: Protect KPI Dashboard sheet with password and specific permissions
Task ID: calc_ps_029
Domain: libreoffice_calc

Creates a KPI Dashboard spreadsheet with data tables, formatting, and a
pivot-table-style summary region in E1:H15. Sheet is UNPROTECTED.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'KPI Dashboard'

    # === Styles ===
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )
    currency_fmt = '$#,##0'
    pct_fmt = '0.0%'
    green_fill = PatternFill(start_color='FFC6EFCE', end_color='FFC6EFCE', fill_type='solid')
    red_fill = PatternFill(start_color='FFFFC7CE', end_color='FFFFC7CE', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFEB9C', end_color='FFFFEB9C', fill_type='solid')

    # === Main data table: A1:D20 ===
    main_headers = ['KPI Metric', 'Q1 Actual', 'Q1 Target', 'Variance']
    for c, h in enumerate(main_headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    kpi_data = [
        ['Revenue', 1250000, 1200000],
        ['Gross Profit', 625000, 600000],
        ['Operating Expenses', 380000, 400000],
        ['Net Income', 245000, 200000],
        ['Customer Acquisition Cost', 185, 200],
        ['Customer Lifetime Value', 2340, 2100],
        ['Monthly Active Users', 45200, 42000],
        ['Churn Rate', 0.032, 0.04],
        ['NPS Score', 72, 65],
        ['Employee Satisfaction', 4.2, 4.0],
        ['Avg Response Time (hrs)', 1.8, 2.5],
        ['Support Ticket Resolution', 0.94, 0.90],
        ['Website Conversion Rate', 0.038, 0.035],
        ['Email Open Rate', 0.245, 0.22],
        ['Social Media Engagement', 0.068, 0.055],
        ['Inventory Turnover', 8.5, 7.5],
        ['Days Sales Outstanding', 32, 35],
        ['Return on Investment', 0.186, 0.15],
        ['Market Share', 0.124, 0.11],
    ]

    for r, (metric, actual, target) in enumerate(kpi_data, 2):
        ws.cell(row=r, column=1, value=metric).border = thin_border
        c_actual = ws.cell(row=r, column=2, value=actual)
        c_actual.border = thin_border
        c_target = ws.cell(row=r, column=3, value=target)
        c_target.border = thin_border

        # Variance column
        if isinstance(actual, float) and actual < 1:
            variance = actual - target
            c_var = ws.cell(row=r, column=4, value=variance)
            c_var.number_format = '0.000'
        elif isinstance(actual, (int, float)) and actual > 1000:
            variance = actual - target
            c_var = ws.cell(row=r, column=4, value=variance)
            c_var.number_format = currency_fmt
        else:
            variance = actual - target
            c_var = ws.cell(row=r, column=4, value=variance)
            c_var.number_format = '0.0'
        c_var.border = thin_border

        # Conditional color on variance
        if metric in ['Operating Expenses', 'Churn Rate', 'Customer Acquisition Cost',
                       'Avg Response Time (hrs)', 'Days Sales Outstanding']:
            # Lower is better for these metrics
            if variance < 0:
                c_var.fill = green_fill
            elif variance > 0:
                c_var.fill = red_fill
            else:
                c_var.fill = yellow_fill
        else:
            if variance > 0:
                c_var.fill = green_fill
            elif variance < 0:
                c_var.fill = red_fill
            else:
                c_var.fill = yellow_fill

        # Number formats for percentages
        if metric in ['Churn Rate', 'Support Ticket Resolution', 'Website Conversion Rate',
                       'Email Open Rate', 'Social Media Engagement', 'Return on Investment',
                       'Market Share']:
            c_actual.number_format = pct_fmt
            c_target.number_format = pct_fmt

        if metric in ['Revenue', 'Gross Profit', 'Operating Expenses', 'Net Income',
                       'Customer Lifetime Value']:
            c_actual.number_format = currency_fmt
            c_target.number_format = currency_fmt

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16

    # === Pivot-table-style summary: E1:H15 ===
    pivot_header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    pivot_headers = ['Category', 'Total Actual', 'Total Target', 'Status']
    for c, h in enumerate(pivot_headers, 5):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = pivot_header_fill
        cell.alignment = header_align
        cell.border = thin_border

    pivot_data = [
        ['Financial', 1740000, 1600000, 'On Track'],
        ['Customer', 47725, 44300, 'On Track'],
        ['Engagement', 0.117, 0.103, 'Exceeding'],
        ['Operations', 33.8, 37.5, 'On Track'],
        ['Growth', 0.186, 0.15, 'Exceeding'],
        ['Marketing', 0.245, 0.22, 'On Track'],
        ['Support', 0.94, 0.90, 'Exceeding'],
        ['Retention', 0.032, 0.04, 'On Track'],
        ['Sales', 8.5, 7.5, 'Exceeding'],
        ['Satisfaction', 4.2, 4.0, 'On Track'],
        ['Digital', 0.038, 0.035, 'On Track'],
        ['Efficiency', 1.8, 2.5, 'Exceeding'],
        ['Brand', 0.124, 0.11, 'On Track'],
        ['Innovation', 0.068, 0.055, 'Exceeding'],
    ]

    for r, (cat, actual, target, status) in enumerate(pivot_data, 2):
        ws.cell(row=r, column=5, value=cat).border = thin_border
        ws.cell(row=r, column=6, value=actual).border = thin_border
        ws.cell(row=r, column=7, value=target).border = thin_border
        c_status = ws.cell(row=r, column=8, value=status)
        c_status.border = thin_border
        c_status.alignment = Alignment(horizontal='center')
        if status == 'Exceeding':
            c_status.fill = green_fill
            c_status.font = Font(color='006100')
        elif status == 'On Track':
            c_status.fill = yellow_fill
            c_status.font = Font(color='9C5700')

    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14

    # Freeze panes at A2 (keep headers visible)
    ws.freeze_panes = 'A2'

    # Sheet is UNPROTECTED (the task is to protect it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
