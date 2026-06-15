"""
Reward Script: Protect 'KPI Dashboard' sheet with password, allow formatting and pivot tables
Task ID: calc_ps_029
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Sheet protection is enabled (sheet=True)
  Component 2 (0.25): Password hash is set (non-None)
  Component 3 (0.20): formatCells is False (users CAN format cells)
  Component 4 (0.20): pivotTables is False (users CAN use pivot tables)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_029'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'KPI Dashboard' sheet must exist
    if 'KPI Dashboard' not in wb.sheetnames:
        print(f"FAIL: Sheet 'KPI Dashboard' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['KPI Dashboard']
    prot = ws.protection

    # Component 1: Sheet protection is enabled (0.35 points)
    # In openpyxl, protection.sheet == True means protection is active.
    # Initial env has sheet=False, golden has sheet=True.
    try:
        if prot.sheet is True:
            print(f"PASS: Component 1 — Sheet protection is enabled (sheet={prot.sheet}) (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Sheet protection not enabled (sheet={prot.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Password hash is set (0.25 points)
    # Initial env has password=None, golden has a password hash.
    # We check that a password hash exists (set_password hashes the password).
    try:
        has_password = (prot.password is not None and prot.password != '')
        has_hash = (prot.hashValue is not None and prot.hashValue != '')
        if has_password or has_hash:
            print(f"PASS: Component 2 — Password is set (password={prot.password}, hashValue={prot.hashValue}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — No password set (password={prot.password}, hashValue={prot.hashValue})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: formatCells is False — users CAN format cells (0.20 points)
    # In openpyxl SheetProtection, formatCells=True means formatting is BLOCKED,
    # formatCells=False means formatting is ALLOWED.
    # Initial env has formatCells=True (blocked by default), golden has formatCells=False (allowed).
    # But we must ensure this component only scores the task change:
    # The change is: protection enabled + formatCells=False (allowed).
    # On initial env, sheet=False so protection is off entirely — formatCells value is irrelevant.
    # We gate on sheet being True so this only passes when protection is active AND formatting is allowed.
    try:
        if prot.sheet is True and prot.formatCells is False:
            print(f"PASS: Component 3 — Format cells allowed under protection (formatCells={prot.formatCells}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — formatCells={prot.formatCells}, sheet={prot.sheet} (need sheet=True, formatCells=False)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: pivotTables is False — users CAN use pivot tables (0.20 points)
    # pivotTables=True means pivot tables are BLOCKED, False means ALLOWED.
    # Initial env has pivotTables=True (blocked by default), golden has pivotTables=False (allowed).
    # Same gating on sheet=True to ensure protection is active.
    try:
        if prot.sheet is True and prot.pivotTables is False:
            print(f"PASS: Component 4 — Pivot tables allowed under protection (pivotTables={prot.pivotTables}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — pivotTables={prot.pivotTables}, sheet={prot.sheet} (need sheet=True, pivotTables=False)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
