"""
Reward Script: JSPS Kakenhi Grant Statistics Table
Task ID: osworld_multi_apps_pdf_stats_table_013
Domain: libreoffice_calc
Scoring:
  Component 1: Correct headers (Year, Applications, Grants, Pass Rate (%)) — 0.3 pts
  Component 2: Correct year column (2019–2023 in rows 2–6) — 0.3 pts
  Component 3: Correct Applications and Grants values for all 5 years — 0.2 pts
  Component 4: Correct Pass Rate values (to 2 decimal places) for all 5 years — 0.2 pts
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_stats_table_013'
FILE_PATH = f'{WORKDIR}/JSPS_KakenA_rates.xlsx'

# Ground truth data extracted from task context / PDFs
# year -> (applications, grants, pass_rate)
EXPECTED_DATA = {
    2019: (2850, 572, 20.07),
    2020: (2780, 548, 19.71),
    2021: (2690, 524, 19.48),
    2022: (2730, 539, 19.74),
    2023: (2810, 563, 20.04),
}

EXPECTED_YEARS = [2019, 2020, 2021, 2022, 2023]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Checks: headers, year sequence, applications/grants values, pass rate values.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the active sheet (task does not specify a specific sheet name)
    ws = wb.active

    # Component 1: Correct headers in row 1 (0.3 points)
    # Headers must be: Year, Applications, Grants, Pass Rate (%)
    try:
        expected_headers = ['Year', 'Applications', 'Grants', 'Pass Rate (%)']
        actual_headers = [
            str(ws.cell(row=1, column=c).value).strip() if ws.cell(row=1, column=c).value is not None else ''
            for c in range(1, 5)
        ]
        if actual_headers == expected_headers:
            print(f"PASS: Component 1 — Headers correct: {actual_headers} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Expected headers {expected_headers}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Year column contains 2019–2023 in rows 2–6 (0.3 points)
    # All 5 years must be present in correct order
    try:
        actual_years = []
        for r in range(2, 7):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                try:
                    actual_years.append(int(val))
                except (ValueError, TypeError):
                    actual_years.append(val)

        if actual_years == EXPECTED_YEARS:
            print(f"PASS: Component 2 — Year column correct: {actual_years} (0.3 pts)")
            total_score += 0.3
        else:
            # Partial check: at least check years are present (any order)
            years_match = set(actual_years) == set(EXPECTED_YEARS)
            if years_match:
                print(f"FAIL: Component 2 — Years present but wrong order. Expected {EXPECTED_YEARS}, found {actual_years}")
            else:
                print(f"FAIL: Component 2 — Expected years {EXPECTED_YEARS}, found {actual_years}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Applications and Grants values correct for all 5 years (0.2 points)
    # Must match expected values from PDF data
    try:
        year_col = 1
        apps_col = 2
        grants_col = 3

        apps_grants_pass = 0
        apps_grants_total = 5

        for r in range(2, 7):
            year_val = ws.cell(row=r, column=year_col).value
            apps_val = ws.cell(row=r, column=apps_col).value
            grants_val = ws.cell(row=r, column=grants_col).value

            try:
                year_int = int(year_val)
            except (ValueError, TypeError):
                print(f"FAIL: Component 3 — Row {r}: Cannot parse year value '{year_val}'")
                continue

            if year_int not in EXPECTED_DATA:
                print(f"FAIL: Component 3 — Row {r}: Unexpected year {year_int}")
                continue

            exp_apps, exp_grants, _ = EXPECTED_DATA[year_int]

            try:
                apps_ok = int(apps_val) == exp_apps
                grants_ok = int(grants_val) == exp_grants
            except (ValueError, TypeError):
                print(f"FAIL: Component 3 — Row {r} (year {year_int}): Non-numeric apps/grants values: apps={apps_val}, grants={grants_val}")
                continue

            if apps_ok and grants_ok:
                apps_grants_pass += 1
            else:
                if not apps_ok:
                    print(f"FAIL: Component 3 — Year {year_int}: Applications expected {exp_apps}, found {apps_val}")
                if not grants_ok:
                    print(f"FAIL: Component 3 — Year {year_int}: Grants expected {exp_grants}, found {grants_val}")

        if apps_grants_pass == apps_grants_total:
            print(f"PASS: Component 3 — All {apps_grants_total} Applications/Grants values correct (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {apps_grants_pass}/{apps_grants_total} Applications/Grants rows correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Pass Rate values correct (to 2 decimal places) for all 5 years (0.2 points)
    # Must match computed pass rates: grants/applications * 100, rounded to 2 decimal places
    try:
        year_col = 1
        rate_col = 4
        TOLERANCE = 0.005  # allow tiny floating-point rounding differences

        rate_pass = 0
        rate_total = 5

        for r in range(2, 7):
            year_val = ws.cell(row=r, column=year_col).value
            rate_val = ws.cell(row=r, column=rate_col).value

            try:
                year_int = int(year_val)
            except (ValueError, TypeError):
                print(f"FAIL: Component 4 — Row {r}: Cannot parse year value '{year_val}'")
                continue

            if year_int not in EXPECTED_DATA:
                print(f"FAIL: Component 4 — Row {r}: Unexpected year {year_int}")
                continue

            _, _, exp_rate = EXPECTED_DATA[year_int]

            try:
                actual_rate = float(rate_val)
            except (ValueError, TypeError):
                print(f"FAIL: Component 4 — Year {year_int}: Non-numeric pass rate value: {rate_val}")
                continue

            if abs(actual_rate - exp_rate) <= TOLERANCE:
                rate_pass += 1
            else:
                print(f"FAIL: Component 4 — Year {year_int}: Pass Rate expected {exp_rate}, found {actual_rate}")

        if rate_pass == rate_total:
            print(f"PASS: Component 4 — All {rate_total} Pass Rate values correct to 2 decimal places (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Only {rate_pass}/{rate_total} Pass Rate values correct")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint: test against canonical artifact path
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
