"""
Reward Script: Sales Team KPI Scorecard
Task ID: calc_wf_045
Domain: libreoffice_calc
Scoring:
  Component 1: Scorecard actuals data grid with cross-sheet references (0.25)
  Component 2: Achievement % section with correct formulas (0.25)
  Component 3: Team average formulas in column G (0.15)
  Component 4: Traffic light conditional formatting on achievement % range (0.15)
  Component 5: Grouped bar chart present on Scorecard (0.10)
  Component 6: Overall team score weighted average formula (0.10)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_045'

REPS = ['Rep_Alice', 'Rep_Bob', 'Rep_Carol', 'Rep_Dan', 'Rep_Eve']
KPIS = ['Calls', 'Meetings', 'Proposals', 'Closed Deals', 'Revenue']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Scorecard sheet must exist
    if 'Scorecard' not in wb.sheetnames:
        print("FAIL: 'Scorecard' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scorecard']

    # ---------------------------------------------------------------
    # Component 1: Scorecard actuals data grid with cross-sheet refs (0.25 points)
    # The Scorecard should pull actual values from each rep sheet via
    # formulas like =Rep_Alice!C2. We check that cells in the actuals
    # area (approximately rows 3-8 or wherever KPIs appear, columns B-F)
    # contain cross-sheet reference formulas.
    # ---------------------------------------------------------------
    try:
        comp1_score = 0.0
        # Find the row that has the actuals header row (contains rep names)
        actuals_header_row = None
        for r in range(1, 25):
            vals_in_row = []
            for c in range(1, 10):
                v = ws.cell(row=r, column=c).value
                if v is not None:
                    vals_in_row.append(str(v))
            # Look for a row containing multiple rep names
            rep_count = sum(1 for rep in REPS if any(rep in v for v in vals_in_row))
            if rep_count >= 3:
                actuals_header_row = r
                break

        if actuals_header_row is None:
            print("FAIL: Component 1 -- Could not find actuals header row with rep names")
        else:
            # Check formulas in the data rows below the header
            cross_ref_count = 0
            total_cells_to_check = 0
            for kpi_offset in range(1, 6):  # 5 KPIs
                data_row = actuals_header_row + kpi_offset
                for col in range(2, 7):  # columns B-F (5 reps)
                    total_cells_to_check += 1
                    cell_val = ws.cell(row=data_row, column=col).value
                    if cell_val is not None and isinstance(cell_val, str) and '!' in cell_val and cell_val.startswith('='):
                        cross_ref_count += 1

            if total_cells_to_check > 0:
                ratio = cross_ref_count / total_cells_to_check
                comp1_score = 0.25 * ratio
                print(f"PASS: Component 1 -- {cross_ref_count}/{total_cells_to_check} cells have cross-sheet refs ({comp1_score:.3f} pts)")
            else:
                print("FAIL: Component 1 -- No data cells found in actuals grid")

        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # ---------------------------------------------------------------
    # Component 2: Achievement % section with actual/target*100 formulas (0.25 points)
    # There should be an achievement percentage section with formulas that
    # divide actual by target (e.g., =Rep_Alice!C2/Rep_Alice!B2*100).
    # ---------------------------------------------------------------
    try:
        comp2_score = 0.0
        # Search for a cell containing "Achievement" (case-insensitive)
        achievement_header_row = None
        for r in range(1, 30):
            v = ws.cell(row=r, column=1).value
            if v and 'achievement' in str(v).lower():
                achievement_header_row = r
                break

        if achievement_header_row is None:
            # Also try to find by looking for rows with division formulas
            for r in range(1, 30):
                v = ws.cell(row=r, column=2).value
                if v and isinstance(v, str) and '/' in v and '*100' in str(v).lower().replace(' ', ''):
                    # Found a row with achievement formula, header is likely r-1 or r-2
                    achievement_header_row = r - 1
                    break

        if achievement_header_row is None:
            print("FAIL: Component 2 -- No achievement percentage section found")
        else:
            # The data rows start 1-2 rows after the header
            # Find the first row after achievement_header_row that has formulas
            ach_data_start = None
            for r in range(achievement_header_row, achievement_header_row + 4):
                v = ws.cell(row=r, column=2).value
                if v and isinstance(v, str) and '/' in v:
                    ach_data_start = r
                    break

            if ach_data_start is None:
                print("FAIL: Component 2 -- Achievement section found but no division formulas")
            else:
                ach_formula_count = 0
                ach_cells_checked = 0
                for kpi_offset in range(5):  # 5 KPIs
                    row = ach_data_start + kpi_offset
                    for col in range(2, 7):  # B-F
                        ach_cells_checked += 1
                        val = ws.cell(row=row, column=col).value
                        if val and isinstance(val, str) and '/' in val and '!' in val:
                            ach_formula_count += 1

                if ach_cells_checked > 0:
                    ratio = ach_formula_count / ach_cells_checked
                    comp2_score = 0.25 * ratio
                    print(f"PASS: Component 2 -- {ach_formula_count}/{ach_cells_checked} achievement % formulas found ({comp2_score:.3f} pts)")
                else:
                    print("FAIL: Component 2 -- No achievement cells checked")

        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # ---------------------------------------------------------------
    # Component 3: Team average formulas in column G (0.15 points)
    # Column G should have AVERAGE formulas aggregating across reps.
    # ---------------------------------------------------------------
    try:
        comp3_score = 0.0
        avg_formula_count = 0
        # Check all rows in column G for AVERAGE formulas
        for r in range(1, 30):
            val = ws.cell(row=r, column=7).value  # column G
            if val and isinstance(val, str) and 'AVERAGE' in val.upper() and val.startswith('='):
                avg_formula_count += 1

        # We expect at least 5 AVERAGE formulas (one per KPI in actuals section)
        # and potentially more in the achievement section
        if avg_formula_count >= 5:
            comp3_score = 0.15
            print(f"PASS: Component 3 -- Found {avg_formula_count} AVERAGE formulas in column G (0.15 pts)")
        elif avg_formula_count >= 3:
            comp3_score = 0.10
            print(f"PARTIAL: Component 3 -- Found {avg_formula_count} AVERAGE formulas in column G (0.10 pts)")
        elif avg_formula_count >= 1:
            comp3_score = 0.05
            print(f"PARTIAL: Component 3 -- Found {avg_formula_count} AVERAGE formulas in column G (0.05 pts)")
        else:
            print("FAIL: Component 3 -- No AVERAGE formulas found in column G")

        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # ---------------------------------------------------------------
    # Component 4: Traffic light conditional formatting (0.15 points)
    # Achievement % cells should have conditional formatting with
    # at least 2-3 rules (green >=100, yellow 80-99, red <80).
    # ---------------------------------------------------------------
    try:
        comp4_score = 0.0
        cf_rules_found = 0
        has_green_rule = False
        has_yellow_rule = False
        has_red_rule = False

        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                cf_rules_found += 1
                rule_type = getattr(rule, 'type', '')
                formulas = getattr(rule, 'formula', []) or []

                # Check for traffic light patterns
                # Green: >= 100
                if rule_type == 'cellIs':
                    op = getattr(rule, 'operator', '')
                    if op == 'greaterThanOrEqual' and any('100' in str(f) for f in formulas):
                        has_green_rule = True
                    # Yellow: between 80 and ~100
                    if op == 'between' and any('80' in str(f) for f in formulas):
                        has_yellow_rule = True
                    # Red: < 80
                    if op == 'lessThan' and any('80' in str(f) for f in formulas):
                        has_red_rule = True

                # Also handle formula-based rules
                if rule_type == 'expression':
                    formula_str = ' '.join(str(f) for f in formulas)
                    if '100' in formula_str:
                        has_green_rule = True
                    if '80' in formula_str:
                        has_yellow_rule = True

        traffic_light_count = sum([has_green_rule, has_yellow_rule, has_red_rule])
        if traffic_light_count >= 3:
            comp4_score = 0.15
            print(f"PASS: Component 4 -- Traffic light CF with {cf_rules_found} rules (green={has_green_rule}, yellow={has_yellow_rule}, red={has_red_rule}) (0.15 pts)")
        elif traffic_light_count >= 2:
            comp4_score = 0.10
            print(f"PARTIAL: Component 4 -- {traffic_light_count}/3 traffic light rules found (0.10 pts)")
        elif cf_rules_found >= 1:
            comp4_score = 0.05
            print(f"PARTIAL: Component 4 -- {cf_rules_found} CF rules found but incomplete traffic light (0.05 pts)")
        else:
            print("FAIL: Component 4 -- No conditional formatting found on Scorecard")

        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # ---------------------------------------------------------------
    # Component 5: Grouped bar chart on Scorecard (0.10 points)
    # There should be at least one bar/column chart with multiple series.
    # ---------------------------------------------------------------
    try:
        comp5_score = 0.0
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_class = chart.__class__.__name__
            is_bar = chart_class in ('BarChart', 'BarChart3D')
            series_count = len(chart.series)

            if is_bar and series_count >= 3:
                comp5_score = 0.10
                print(f"PASS: Component 5 -- Bar chart found with {series_count} series (0.10 pts)")
            elif is_bar:
                comp5_score = 0.05
                print(f"PARTIAL: Component 5 -- Bar chart found but only {series_count} series (0.05 pts)")
            elif series_count >= 3:
                comp5_score = 0.05
                print(f"PARTIAL: Component 5 -- Chart found ({chart_class}) with {series_count} series but not a bar chart (0.05 pts)")
            else:
                comp5_score = 0.03
                print(f"PARTIAL: Component 5 -- Chart exists but is {chart_class} with {series_count} series (0.03 pts)")
        else:
            print("FAIL: Component 5 -- No chart found on Scorecard sheet")

        total_score += comp5_score
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # ---------------------------------------------------------------
    # Component 6: Overall team score with weighted average (0.10 points)
    # There should be a cell with a weighted average formula combining
    # the team averages with weights.
    # ---------------------------------------------------------------
    try:
        comp6_score = 0.0
        weighted_formula_found = False

        # Search for a cell containing a weighted formula or "Overall" label
        for r in range(1, 30):
            label = ws.cell(row=r, column=1).value
            if label and 'overall' in str(label).lower():
                # Check the formula in the adjacent cell (column B or nearby)
                for c in range(2, 10):
                    val = ws.cell(row=r, column=c).value
                    if val and isinstance(val, str) and val.startswith('='):
                        # A formula next to "Overall" label -- check if it references
                        # team average cells with multiplication (weighting)
                        if '*' in val and ('+' in val or 'SUM' in val.upper()):
                            weighted_formula_found = True
                            break
                        # Also accept SUMPRODUCT or similar
                        if 'SUMPRODUCT' in val.upper():
                            weighted_formula_found = True
                            break
                break

        if weighted_formula_found:
            comp6_score = 0.10
            print(f"PASS: Component 6 -- Weighted overall team score formula found (0.10 pts)")
        else:
            # Fallback: look for any cell with weighted average pattern anywhere
            for r in range(1, 30):
                for c in range(1, 10):
                    val = ws.cell(row=r, column=c).value
                    if val and isinstance(val, str) and val.startswith('=') and '*' in val:
                        # Count multiplication operations -- weighted formulas have multiple
                        mult_count = val.count('*')
                        plus_count = val.count('+')
                        if mult_count >= 3 and plus_count >= 2:
                            weighted_formula_found = True
                            comp6_score = 0.10
                            print(f"PASS: Component 6 -- Weighted formula found at ({r},{c}) (0.10 pts)")
                            break
                if weighted_formula_found:
                    break

            if not weighted_formula_found:
                print("FAIL: Component 6 -- No weighted overall team score formula found")

        total_score += comp6_score
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for unsaved GUI edits
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state()

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
