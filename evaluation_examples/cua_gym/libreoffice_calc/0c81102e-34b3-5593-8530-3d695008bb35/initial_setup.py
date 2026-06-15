"""
Initial Setup: KPI Scorecard for Sales Team
Task ID: calc_wf_045
Domain: libreoffice_calc

Creates 5 rep sheets with monthly targets and actuals for 5 KPIs,
plus a blank Scorecard sheet ready for the agent to populate.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # KPI definitions with realistic sales targets and actuals per rep
    kpis = ['Calls', 'Meetings', 'Proposals', 'Closed Deals', 'Revenue']

    rep_data = {
        'Rep_Alice': {
            'Calls':        (150, 172),
            'Meetings':     (30,  28),
            'Proposals':    (15,  18),
            'Closed Deals': (8,   9),
            'Revenue':      (120000, 134500),
        },
        'Rep_Bob': {
            'Calls':        (150, 131),
            'Meetings':     (30,  24),
            'Proposals':    (15,  11),
            'Closed Deals': (8,   6),
            'Revenue':      (120000, 95800),
        },
        'Rep_Carol': {
            'Calls':        (150, 158),
            'Meetings':     (30,  33),
            'Proposals':    (15,  16),
            'Closed Deals': (8,   10),
            'Revenue':      (120000, 142300),
        },
        'Rep_Dan': {
            'Calls':        (150, 119),
            'Meetings':     (30,  22),
            'Proposals':    (15,  12),
            'Closed Deals': (8,   5),
            'Revenue':      (120000, 78400),
        },
        'Rep_Eve': {
            'Calls':        (150, 145),
            'Meetings':     (30,  30),
            'Proposals':    (15,  14),
            'Closed Deals': (8,   8),
            'Revenue':      (120000, 118900),
        },
    }

    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    rep_names = ['Rep_Alice', 'Rep_Bob', 'Rep_Carol', 'Rep_Dan', 'Rep_Eve']

    # --- Create individual rep sheets ---
    first_sheet = True
    for rep_name in rep_names:
        if first_sheet:
            ws = wb.active
            ws.title = rep_name
            first_sheet = False
        else:
            ws = wb.create_sheet(rep_name)

        headers = ['KPI', 'Monthly Target', 'Actual']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 15

        for r, kpi in enumerate(kpis, 2):
            target, actual = rep_data[rep_name][kpi]
            ws.cell(row=r, column=1, value=kpi).border = thin_border
            ws.cell(row=r, column=1).alignment = Alignment(horizontal='left')

            target_cell = ws.cell(row=r, column=2, value=target)
            target_cell.border = thin_border
            target_cell.alignment = Alignment(horizontal='center')

            actual_cell = ws.cell(row=r, column=3, value=actual)
            actual_cell.border = thin_border
            actual_cell.alignment = Alignment(horizontal='center')

            # Format Revenue row as currency
            if kpi == 'Revenue':
                target_cell.number_format = '$#,##0'
                actual_cell.number_format = '$#,##0'

    # --- Create blank Scorecard sheet (no formulas, no chart, no formatting) ---
    ws_sc = wb.create_sheet('Scorecard')
    ws_sc.column_dimensions['A'].width = 18

    # Just a simple header label so the agent knows where to work
    ws_sc['A1'] = 'Sales Team KPI Scorecard'
    ws_sc['A1'].font = Font(name='Calibri', size=14, bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
