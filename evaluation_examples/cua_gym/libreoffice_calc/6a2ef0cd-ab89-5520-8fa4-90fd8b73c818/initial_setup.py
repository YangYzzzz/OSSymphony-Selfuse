"""
Initial Setup: Department budget spreadsheet without average row and without charts
Task ID: osworld_calc_multi_chart_computed_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Budget ---
    ws = wb.active
    ws.title = "Budget"

    # Headers row
    headers = ["Department", "Q1", "Q2", "Q3", "Q4"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Department budget data (5 departments, realistic business data)
    data = [
        ["Engineering",    145000, 152000, 148000, 163000],
        ["Marketing",       98000,  112000, 105000, 121000],
        ["Sales",          175000,  183000, 190000, 205000],
        ["Human Resources", 62000,   65000,  63000,  68000],
        ["Operations",     135000,  141000, 138000, 150000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 14

    # NOTE: No average row (row 7 is empty) — task requires agent to add it
    # NOTE: No charts — task requires agent to create two charts

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
