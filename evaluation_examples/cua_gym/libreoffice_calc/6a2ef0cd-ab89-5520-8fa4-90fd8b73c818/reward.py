"""
Reward Script: Add average row and create two charts for department budget analysis
Task ID: osworld_calc_multi_chart_computed_005
Domain: libreoffice_calc
Scoring:
  - Component 1: Average row exists with AVERAGE formulas (0.35 pts)
  - Component 2: Bar chart exists with title "Department Budget Comparison" (0.35 pts)
  - Component 3: Line chart exists with title "Budget Trend by Quarter" (0.30 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_005'


def extract_chart_title(chart):
    """Extract chart title text from openpyxl chart object."""
    try:
        if chart.title is None:
            return None
        t = chart.title
        # Extract from RichText paragraphs
        if hasattr(t, 'tx') and t.tx is not None:
            tx = t.tx
            if hasattr(tx, 'rich') and tx.rich is not None:
                for para in tx.rich.p:
                    if hasattr(para, 'r'):
                        for r in para.r:
                            if r.t:
                                return r.t
        # Fallback: try strRef
        if hasattr(t, 'tx') and t.tx is not None:
            tx = t.tx
            if hasattr(tx, 'strRef') and tx.strRef is not None:
                return tx.strRef
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Ensure 'Budget' sheet exists (precondition gate)
    if 'Budget' not in wb.sheetnames:
        print("CRITICAL: 'Budget' sheet not found in workbook.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget']

    # Component 1: Average row with AVERAGE formulas (0.35 points)
    # The task requires an average row below the data (row 7 for 5 data rows + 1 header).
    # A7 should contain 'Average' and B7:E7 should contain =AVERAGE(...) formulas.
    try:
        avg_label = ws.cell(row=7, column=1).value
        avg_label_ok = avg_label is not None and str(avg_label).strip().lower() == 'average'

        # Check AVERAGE formulas in B7:E7
        formula_count = 0
        for col in range(2, 6):  # columns B, C, D, E
            cell_val = ws.cell(row=7, column=col).value
            if cell_val is not None and isinstance(cell_val, str) and 'AVERAGE' in cell_val.upper():
                formula_count += 1

        if avg_label_ok and formula_count == 4:
            print(f"PASS: Component 1 — Average row with label and 4 AVERAGE formulas found (0.35 pts)")
            total_score += 0.35
        elif avg_label_ok and formula_count >= 2:
            print(f"PASS (partial): Component 1 — Average label OK, but only {formula_count}/4 AVERAGE formulas found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Average row missing or incomplete. "
                  f"label={avg_label!r}, AVERAGE formula count={formula_count}/4")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Bar chart titled "Department Budget Comparison" (0.35 points)
    # The task requires a bar chart comparing department budgets.
    try:
        charts = ws._charts
        bar_chart_found = False
        bar_title_correct = False

        for chart in charts:
            chart_type = type(chart).__name__
            if chart_type == 'BarChart':
                bar_chart_found = True
                title_text = extract_chart_title(chart)
                if title_text and 'Department Budget Comparison' in title_text:
                    bar_title_correct = True
                    break

        if bar_chart_found and bar_title_correct:
            print(f"PASS: Component 2 — Bar chart with title 'Department Budget Comparison' found (0.35 pts)")
            total_score += 0.35
        elif bar_chart_found:
            title_text = extract_chart_title(charts[[type(c).__name__ for c in charts].index('BarChart')])
            print(f"FAIL: Component 2 — Bar chart found but title incorrect. Found: {title_text!r}")
        else:
            print(f"FAIL: Component 2 — No BarChart found among {len(charts)} chart(s).")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Line chart titled "Budget Trend by Quarter" (0.30 points)
    # The task requires a line chart showing budget trends by quarter.
    try:
        charts = ws._charts
        line_chart_found = False
        line_title_correct = False

        for chart in charts:
            chart_type = type(chart).__name__
            if chart_type == 'LineChart':
                line_chart_found = True
                title_text = extract_chart_title(chart)
                if title_text and 'Budget Trend by Quarter' in title_text:
                    line_title_correct = True
                    break

        if line_chart_found and line_title_correct:
            print(f"PASS: Component 3 — Line chart with title 'Budget Trend by Quarter' found (0.30 pts)")
            total_score += 0.30
        elif line_chart_found:
            title_text = None
            for chart in charts:
                if type(chart).__name__ == 'LineChart':
                    title_text = extract_chart_title(chart)
                    break
            print(f"FAIL: Component 3 — Line chart found but title incorrect. Found: {title_text!r}")
        else:
            print(f"FAIL: Component 3 — No LineChart found among {len(charts)} chart(s).")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
