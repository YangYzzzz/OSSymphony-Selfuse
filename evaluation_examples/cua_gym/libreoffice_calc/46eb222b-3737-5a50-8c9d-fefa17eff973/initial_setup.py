"""
Initial Setup: Operating Lease Amortization (ASC 842)
Task ID: calc_fin_operating_lease_078
Domain: libreoffice_calc

Creates the pre-task state:
- Sheet 'LeaseCalc' with lease parameters in B1:B3
- B4 empty (PV to be calculated by agent)
- Row 6 headers (Period, Payment, Interest Expense, Liability Reduction, Ending Liability)
- A7:A66 filled with period numbers 1-60
- B7:E66 empty (amortization table to be filled by agent)
- No bold formatting on row 6, no freeze panes
- No currency formatting
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_operating_lease_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: LeaseCalc ---
    ws = wb.active
    ws.title = 'LeaseCalc'

    # ---- Lease parameters (B1:B3) ----
    ws['A1'] = 'Incremental Borrowing Rate (Annual)'
    ws['B1'] = 0.045   # 4.5%

    ws['A2'] = 'Lease Term (Years)'
    ws['B2'] = 5

    ws['A3'] = 'Annual Lease Payment ($)'
    ws['B3'] = 120000

    ws['A4'] = 'Present Value of Lease Liability'
    # B4 intentionally left empty — agent must calculate it

    # ---- Column headers (Row 6) ----
    # NOT bold, NOT frozen — those are part of the task requirement
    headers = {
        'A6': 'Period',
        'B6': 'Payment',
        'C6': 'Interest Expense',
        'D6': 'Liability Reduction',
        'E6': 'Ending Liability',
    }
    for coord, label in headers.items():
        ws[coord] = label

    # ---- Period numbers A7:A66 (60 monthly periods) ----
    for month in range(1, 61):
        ws.cell(row=6 + month, column=1, value=month)

    # B7:E66 intentionally left empty — agent must fill in the amortization table

    # ---- Column widths for readability ----
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
