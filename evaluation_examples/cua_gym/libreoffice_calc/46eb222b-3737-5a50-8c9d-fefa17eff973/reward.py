"""
Reward Script: Calculate present value of operating lease obligations under ASC 842
Task ID: calc_fin_operating_lease_078
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): B4 PV formula present — the lease liability/ROU asset present value
  Component 2 (0.35): Amortization table formulas in B7:E66 (payment, interest, reduction, ending liability)
  Component 3 (0.15): Currency formatting on B4 and B7:E66
  Component 4 (0.15): Row 6 headers are bold and freeze panes set at A7
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_operating_lease_078'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'LeaseCalc' not in wb.sheetnames:
        print("CRITICAL: Sheet 'LeaseCalc' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['LeaseCalc']

    # Component 1: B4 contains a PV formula for the present value of lease liability (0.35 points)
    # This is the core ROU asset/initial lease liability calculation
    # Initial file: B4 is empty. Golden file: B4 = '=PV(B1/12,B2*12,-B3/12)'
    try:
        b4_val = ws.cell(row=4, column=2).value
        if b4_val is not None and isinstance(b4_val, str):
            normalized = b4_val.upper().replace(' ', '')
            # Must contain PV( function call
            if 'PV(' in normalized or '=PV(' in normalized:
                # Also verify it references B1 (rate), B2 (nper), B3 (pmt) in some form
                if 'B1' in normalized and 'B2' in normalized and 'B3' in normalized:
                    print(f"PASS: Component 1 — B4 contains PV formula referencing rate/term/payment: {repr(b4_val)} (0.35 pts)")
                    total_score += 0.35
                else:
                    print(f"FAIL: Component 1 — B4 has PV formula but doesn't reference B1/B2/B3: {repr(b4_val)}")
            else:
                print(f"FAIL: Component 1 — B4 does not contain PV formula, found: {repr(b4_val)}")
        else:
            print(f"FAIL: Component 1 — B4 is empty or not a formula string, found: {repr(b4_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check B4: {e}")

    # Component 2: Amortization table formulas in rows 7-66 (0.35 points)
    # Verify that B, C, D, E columns in rows 7-66 contain formulas (not empty)
    # Initial file: all these cells are empty. Golden: all contain formulas.
    # We check: payment formula in column B, interest in C, liability reduction in D, ending liability in E
    try:
        b_formulas_ok = 0
        c_formulas_ok = 0
        d_formulas_ok = 0
        e_formulas_ok = 0
        rows_to_check = range(7, 67)  # rows 7-66 = 60 rows

        for row in rows_to_check:
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value
            d_val = ws.cell(row=row, column=4).value
            e_val = ws.cell(row=row, column=5).value

            if b_val is not None and isinstance(b_val, str) and b_val.startswith('='):
                b_formulas_ok += 1
            if c_val is not None and isinstance(c_val, str) and c_val.startswith('='):
                c_formulas_ok += 1
            if d_val is not None and isinstance(d_val, str) and d_val.startswith('='):
                d_formulas_ok += 1
            if e_val is not None and isinstance(e_val, str) and e_val.startswith('='):
                e_formulas_ok += 1

        total_rows = len(rows_to_check)  # 60
        # All 4 columns need formulas in all 60 rows
        b_ok = b_formulas_ok == total_rows
        c_ok = c_formulas_ok == total_rows
        d_ok = d_formulas_ok == total_rows
        e_ok = e_formulas_ok == total_rows

        columns_ok = sum([b_ok, c_ok, d_ok, e_ok])

        if columns_ok == 4:
            # All 4 columns have complete formulas — verify first/last formula structure
            # Check row 7: B7 should reference B3, C7 should reference B4 and B1
            b7 = ws.cell(row=7, column=2).value
            c7 = ws.cell(row=7, column=3).value
            d7 = ws.cell(row=7, column=4).value
            e7 = ws.cell(row=7, column=5).value

            b7_ok = b7 and 'B3' in b7.upper() or (b7 and 'B$3' in b7.upper())
            c7_ok = c7 and 'B4' in c7.upper() and 'B1' in c7.upper() or (c7 and 'B$1' in c7.upper())
            d7_ok = d7 and 'B7' in d7.upper() and 'C7' in d7.upper()
            e7_ok = e7 and 'B4' in e7.upper() and 'D7' in e7.upper()

            # Check row 8: C8 should reference E7 (previous ending liability)
            c8 = ws.cell(row=8, column=3).value
            c8_ok = c8 and 'E7' in c8.upper()

            structure_checks = sum([bool(b7_ok), bool(c7_ok), bool(d7_ok), bool(e7_ok), bool(c8_ok)])
            if structure_checks >= 4:
                print(f"PASS: Component 2 — All 60 rows have amortization formulas in B-E, correct structure (0.35 pts)")
                print(f"  B7={repr(b7)}, C7={repr(c7)}, D7={repr(d7)}, E7={repr(e7)}, C8={repr(c8)}")
                total_score += 0.35
            else:
                print(f"FAIL: Component 2 — Formulas present but structure incorrect ({structure_checks}/5 checks passed)")
                print(f"  B7={repr(b7)}, C7={repr(c7)}, D7={repr(d7)}, E7={repr(e7)}, C8={repr(c8)}")
        else:
            print(f"FAIL: Component 2 — Only {columns_ok}/4 columns fully populated: B={b_formulas_ok}, C={c_formulas_ok}, D={d_formulas_ok}, E={e_formulas_ok} (need 60 each)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Currency formatting on B4 and amortization range B7:E66 (0.15 points)
    # Initial file: no currency formatting. Golden: '$#,##0.00' on these cells.
    try:
        currency_formats = ['$#,##0.00', '#,##0.00', '$#,##0', '"$"#,##0.00']

        def is_currency(fmt):
            if fmt is None:
                return False
            return '$' in fmt or (fmt in currency_formats)

        b4_fmt = ws.cell(row=4, column=2).number_format
        b4_currency = is_currency(b4_fmt)

        # Sample check: check B7, C7, D7, E7 and last row E66
        sample_cells = [
            ws.cell(row=7, column=2),   # B7
            ws.cell(row=7, column=3),   # C7
            ws.cell(row=7, column=4),   # D7
            ws.cell(row=7, column=5),   # E7
            ws.cell(row=66, column=5),  # E66
        ]
        currency_count = sum(1 for c in sample_cells if is_currency(c.number_format))

        if b4_currency and currency_count >= 4:
            print(f"PASS: Component 3 — B4 and amortization table cells have currency formatting (0.15 pts)")
            print(f"  B4 format: {repr(b4_fmt)}, B7 format: {repr(ws.cell(row=7, column=2).number_format)}")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Currency formatting incomplete: B4={repr(b4_fmt)}({b4_currency}), sample cells: {currency_count}/5")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row 6 headers bold AND freeze panes at A7 (0.15 points)
    # Initial file: row 6 not bold, no freeze. Golden: bold, frozen.
    try:
        # Check row 6 bold (at least the first 3 header cells must be bold)
        header_cells_bold = 0
        for col in range(1, 6):
            cell = ws.cell(row=6, column=col)
            if cell.font and cell.font.bold:
                header_cells_bold += 1

        headers_bold = header_cells_bold >= 3

        # Check freeze panes at A7 (row 6 frozen = freeze at A7)
        freeze = ws.freeze_panes
        # A7 means rows 1-6 are frozen
        freeze_ok = freeze == 'A7'

        if headers_bold and freeze_ok:
            print(f"PASS: Component 4 — Row 6 headers are bold ({header_cells_bold}/5) and freeze panes at A7 (0.15 pts)")
            total_score += 0.15
        elif headers_bold and not freeze_ok:
            print(f"FAIL: Component 4 — Row 6 bold OK but freeze_panes={repr(freeze)} (expected 'A7')")
        elif not headers_bold and freeze_ok:
            print(f"FAIL: Component 4 — Freeze panes OK but row 6 headers only {header_cells_bold}/5 bold")
        else:
            print(f"FAIL: Component 4 — Row 6 bold {header_cells_bold}/5, freeze_panes={repr(freeze)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
