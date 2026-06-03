"""
Initial Setup: Create expense tracking spreadsheet with 200 rows of department expenses.
Task ID: calc_pivot_018
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    departments = ['Engineering', 'Marketing', 'Sales', 'HR', 'Finance']
    expense_types = ['Travel', 'Meals', 'Equipment', 'Software', 'Training']

    # We need 200 rows with specific constraints:
    # - Engineering/Equipment SUM = 18500, count = 12 (AVG ~1542)
    # - Grand total SUM = 245000
    #
    # Strategy: pre-assign department/expense_type combos, then set amounts
    # to hit targets. 5 departments x 5 types = 25 combos.
    # Distribute 200 rows across combos.

    # First, build all 200 (dept, type) assignments
    # We want ~8 rows per combo (200/25=8), but Engineering/Equipment = 12
    rows = []

    # Assign Engineering/Equipment: 12 rows
    eng_equip_count = 12
    eng_equip_amounts = [
        1800, 1500, 1650, 1200, 1400, 1550, 1700, 1350, 1450, 1600, 1500, 1800
    ]
    # Sum = 18500? Let's check: 1800+1500+1650+1200+1400+1550+1700+1350+1450+1600+1500+1800
    # = 18500. Yes!
    for amt in eng_equip_amounts:
        rows.append(('Engineering', 'Equipment', amt))

    # Now we need 188 more rows, total sum of those = 245000 - 18500 = 226500
    # 24 remaining combos, distribute ~188/24 ~ 7-8 rows each
    remaining_combos = []
    for d in departments:
        for t in expense_types:
            if d == 'Engineering' and t == 'Equipment':
                continue
            remaining_combos.append((d, t))

    # Assign row counts: 24 combos, 188 rows
    # 188 / 24 = 7 remainder 20, so 20 combos get 8 rows, 4 get 7
    combo_counts = {}
    base = 188 // 24  # 7
    extra = 188 % 24  # 20
    for i, combo in enumerate(remaining_combos):
        combo_counts[combo] = base + (1 if i < extra else 0)

    # Target sum for remaining = 226500
    # Average per row = 226500 / 188 = 1204.79
    # Generate amounts per combo to sum to ~226500
    remaining_target = 226500
    remaining_rows_count = 188

    # Generate amounts for each combo
    combo_amounts = {}
    generated_sum = 0
    generated_count = 0

    for i, (combo, count) in enumerate(combo_counts.items()):
        if i < len(combo_counts) - 1:
            # Generate random amounts for this combo
            amounts = []
            avg_target = remaining_target / (remaining_rows_count - generated_count) if (remaining_rows_count - generated_count) > 0 else 1200
            for j in range(count):
                amt = random.randint(int(avg_target * 0.4), int(avg_target * 1.8))
                # Round to nearest 50
                amt = round(amt / 50) * 50
                if amt < 100:
                    amt = 100
                amounts.append(amt)
            combo_amounts[combo] = amounts
            generated_sum += sum(amounts)
            generated_count += count
        else:
            # Last combo: adjust to hit exact target
            needed = remaining_target - generated_sum
            count_left = count
            amounts = []
            for j in range(count_left - 1):
                avg_needed = needed / (count_left - j)
                amt = int(avg_needed + random.randint(-200, 200))
                amt = round(amt / 50) * 50
                if amt < 100:
                    amt = 100
                amounts.append(amt)
                needed -= amt
            # Last value gets the remainder
            last_amt = needed
            if last_amt < 100:
                # Redistribute: reduce previous amounts
                deficit = 100 - last_amt
                last_amt = 100
                # Reduce the first amount
                amounts[0] -= deficit
                if amounts[0] < 100:
                    amounts[0] = 100
            amounts.append(int(last_amt))
            combo_amounts[combo] = amounts
            generated_sum += sum(amounts)

    # Add remaining rows
    for combo, amounts in combo_amounts.items():
        for amt in amounts:
            rows.append((combo[0], combo[1], amt))

    # Adjust total to exactly 245000
    current_total = 18500 + sum(sum(a) for a in combo_amounts.values())
    diff = 245000 - current_total
    # Apply diff to the last non-engineering-equipment row
    if diff != 0:
        for i in range(len(rows) - 1, -1, -1):
            if not (rows[i][0] == 'Engineering' and rows[i][1] == 'Equipment'):
                d, t, a = rows[i]
                rows[i] = (d, t, a + diff)
                break

    # Verify
    total = sum(r[2] for r in rows)
    eng_eq = sum(r[2] for r in rows if r[0] == 'Engineering' and r[1] == 'Equipment')
    assert total == 245000, f"Total mismatch: {total}"
    assert eng_eq == 18500, f"Eng/Equipment mismatch: {eng_eq}"
    assert len(rows) == 200, f"Row count mismatch: {len(rows)}"

    # Shuffle rows
    random.shuffle(rows)

    # Generate dates spanning 2025
    base_date = datetime(2025, 1, 5)
    dates = []
    for _ in range(200):
        offset = random.randint(0, 350)
        dates.append(base_date + timedelta(days=offset))
    dates.sort()

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'DeptExpenses'

    # Headers
    headers = ['ExpID', 'Date', 'Department', 'ExpenseType', 'Amount']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        bottom=Side(style="thin", color="000000")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows
    for i, (dept, etype, amount) in enumerate(rows):
        row_num = i + 2
        ws.cell(row=row_num, column=1, value=i + 1)  # ExpID
        ws.cell(row=row_num, column=2, value=dates[i].strftime('%Y-%m-%d'))  # Date
        ws.cell(row=row_num, column=3, value=dept)
        ws.cell(row=row_num, column=4, value=etype)
        amt_cell = ws.cell(row=row_num, column=5, value=amount)
        amt_cell.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total rows: {len(rows)}, Total amount: {total}')
    print(f'Engineering/Equipment: count={eng_eq_count}, sum={eng_eq}')

    # Launch LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

# Track eng/equip count for reporting
eng_eq_count = 12
create_initial()
