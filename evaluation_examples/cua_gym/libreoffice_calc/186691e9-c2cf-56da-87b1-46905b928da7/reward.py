"""
Reward Script: Pivot table breaking down expenses by department and expense type
Task ID: calc_pivot_018
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - PivotTable sheet exists in workbook
  Component 2 (0.20) - SUM section structure: Department rows + ExpenseType columns
  Component 3 (0.20) - AVERAGE section structure: Department rows + ExpenseType AVG columns
  Component 4 (0.20) - Key SUM values correct (Engineering/Equipment=18500, Grand Total=245000)
  Component 5 (0.20) - Key AVG values correct (Engineering/Equipment~1542)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_018'

# Expected departments and expense types
DEPARTMENTS = {'Engineering', 'Marketing', 'Sales', 'HR', 'Finance'}
EXPENSE_TYPES = {'Travel', 'Meals', 'Equipment', 'Software', 'Training'}


def find_pivot_sheet(wb):
    """Find the pivot table sheet (any sheet other than DeptExpenses)."""
    for sn in wb.sheetnames:
        if sn != 'DeptExpenses':
            ws = wb[sn]
            # Verify it has some pivot-like content (department names)
            all_vals = set()
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1, values_only=True):
                if row[0]:
                    all_vals.add(str(row[0]).strip())
            dept_overlap = DEPARTMENTS & all_vals
            if len(dept_overlap) >= 3:
                return ws
    return None


def find_header_row(ws, keyword):
    """Find a row containing column headers with keyword (e.g. 'SUM' or 'AVG')."""
    for r in range(1, ws.max_row + 1):
        matches = 0
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val and keyword.upper() in str(val).upper():
                matches += 1
        if matches >= 3:  # At least 3 columns with the keyword
            return r
    return None


def parse_pivot_section(ws, header_row):
    """Parse a pivot section starting at header_row.
    Returns dict: {dept: {expense_type: value, ..., 'Grand Total': value}, ...}
    """
    if header_row is None:
        return {}

    # Parse column headers
    col_map = {}  # col_idx -> expense_type_name
    for c in range(2, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val:
            col_map[c] = str(val).strip()

    # Parse data rows below header
    result = {}
    for r in range(header_row + 1, ws.max_row + 1):
        dept = ws.cell(row=r, column=1).value
        if dept is None:
            break
        dept = str(dept).strip()
        row_data = {}
        for c, col_name in col_map.items():
            cell_val = ws.cell(row=r, column=c).value
            row_data[col_name] = cell_val
        result[dept] = row_data

    return result


def get_value_for_type(row_data, type_keyword):
    """Find a value in row_data where the column name contains type_keyword."""
    for col_name, val in row_data.items():
        if type_keyword.upper() in col_name.upper():
            return val
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.2 points)
    # This is the primary task-introduced change — initial has only DeptExpenses
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — No pivot table sheet found with department data")
            # No point continuing if no pivot sheet
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: SUM section structure (0.2 points)
    # Pivot table must have a section with SUM aggregation for each expense type
    try:
        sum_header = find_header_row(pivot_ws, 'SUM')
        if sum_header is not None:
            sum_data = parse_pivot_section(pivot_ws, sum_header)
            # Check departments are present as rows
            found_depts = DEPARTMENTS & set(sum_data.keys())
            if len(found_depts) >= 4:
                # Check expense type columns exist
                if sum_data and len(next(iter(sum_data.values()))) >= 5:
                    print(f"PASS: Component 2 — SUM section has {len(found_depts)} departments and expense type columns (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 2 — SUM section has too few columns")
            else:
                print(f"FAIL: Component 2 — SUM section has only {len(found_depts)} departments, expected >= 4")
        else:
            print("FAIL: Component 2 — No SUM header row found in pivot sheet")
            sum_data = {}
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")
        sum_data = {}

    # Component 3: AVERAGE section structure (0.2 points)
    # Pivot table must also have an AVERAGE aggregation section
    try:
        avg_header = find_header_row(pivot_ws, 'AVG')
        if avg_header is not None:
            avg_data = parse_pivot_section(pivot_ws, avg_header)
            found_depts = DEPARTMENTS & set(avg_data.keys())
            if len(found_depts) >= 4:
                if avg_data and len(next(iter(avg_data.values()))) >= 5:
                    print(f"PASS: Component 3 — AVG section has {len(found_depts)} departments and expense type columns (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 — AVG section has too few columns")
            else:
                print(f"FAIL: Component 3 — AVG section has only {len(found_depts)} departments, expected >= 4")
        else:
            print("FAIL: Component 3 — No AVG/AVERAGE header row found in pivot sheet")
            avg_data = {}
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")
        avg_data = {}

    # Component 4: Key SUM values (0.2 points)
    # Engineering/Equipment SUM = 18500, Grand Total SUM = 245000
    try:
        sub_score = 0.0
        if sum_data and 'Engineering' in sum_data:
            eng_equip = get_value_for_type(sum_data['Engineering'], 'Equipment')
            if eng_equip is not None:
                try:
                    if abs(float(eng_equip) - 18500) < 1:
                        print(f"PASS: Component 4a — Engineering/Equipment SUM = {eng_equip} (expected 18500)")
                        sub_score += 0.1
                    else:
                        print(f"FAIL: Component 4a — Engineering/Equipment SUM = {eng_equip}, expected 18500")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4a — Engineering/Equipment SUM not numeric: {eng_equip}")
            else:
                print("FAIL: Component 4a — Engineering/Equipment SUM not found")
        else:
            print("FAIL: Component 4a — Engineering not in SUM data")

        if sum_data and 'Grand Total' in sum_data:
            grand_total = get_value_for_type(sum_data['Grand Total'], 'Grand Total')
            if grand_total is None:
                # Try last column
                gt_row = sum_data['Grand Total']
                for col_name, val in gt_row.items():
                    if 'grand' in col_name.lower() or 'total' in col_name.lower():
                        grand_total = val
                        break
            if grand_total is not None:
                try:
                    if abs(float(grand_total) - 245000) < 1:
                        print(f"PASS: Component 4b — Grand Total SUM = {grand_total} (expected 245000)")
                        sub_score += 0.1
                    else:
                        print(f"FAIL: Component 4b — Grand Total SUM = {grand_total}, expected 245000")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 4b — Grand Total SUM not numeric: {grand_total}")
            else:
                print("FAIL: Component 4b — Grand Total SUM column not found")
        else:
            print("FAIL: Component 4b — Grand Total row not in SUM data")

        if sub_score > 0:
            total_score += sub_score
            print(f"PASS: Component 4 — Key SUM values verified ({sub_score} pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Key AVG values (0.2 points)
    # Engineering/Equipment AVG ~ 1542 (actual 1541.67)
    try:
        sub_score = 0.0
        if avg_data and 'Engineering' in avg_data:
            eng_equip_avg = get_value_for_type(avg_data['Engineering'], 'Equipment')
            if eng_equip_avg is not None:
                try:
                    val = float(eng_equip_avg)
                    # Allow tolerance: context says 1542, actual is 1541.67
                    if abs(val - 1541.67) < 5:
                        print(f"PASS: Component 5a — Engineering/Equipment AVG = {eng_equip_avg} (expected ~1542)")
                        sub_score += 0.1
                    else:
                        print(f"FAIL: Component 5a — Engineering/Equipment AVG = {eng_equip_avg}, expected ~1542")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 5a — Engineering/Equipment AVG not numeric: {eng_equip_avg}")
            else:
                print("FAIL: Component 5a — Engineering/Equipment AVG not found")
        else:
            print("FAIL: Component 5a — Engineering not in AVG data")

        # Also check Grand Total AVG exists and is reasonable
        if avg_data and 'Grand Total' in avg_data:
            gt_avg = get_value_for_type(avg_data['Grand Total'], 'Grand Total')
            if gt_avg is None:
                gt_row = avg_data['Grand Total']
                for col_name, val in gt_row.items():
                    if 'grand' in col_name.lower() or 'total' in col_name.lower():
                        gt_avg = val
                        break
            if gt_avg is not None:
                try:
                    val = float(gt_avg)
                    # Grand total AVG = 245000 / 200 = 1225
                    if abs(val - 1225) < 5:
                        print(f"PASS: Component 5b — Grand Total AVG = {gt_avg} (expected ~1225)")
                        sub_score += 0.1
                    else:
                        print(f"FAIL: Component 5b — Grand Total AVG = {gt_avg}, expected ~1225")
                except (ValueError, TypeError):
                    print(f"FAIL: Component 5b — Grand Total AVG not numeric: {gt_avg}")
            else:
                print("FAIL: Component 5b — Grand Total AVG column not found")
        else:
            print("FAIL: Component 5b — Grand Total row not in AVG data")

        if sub_score > 0:
            total_score += sub_score
            print(f"PASS: Component 5 — Key AVG values verified ({sub_score} pts)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
