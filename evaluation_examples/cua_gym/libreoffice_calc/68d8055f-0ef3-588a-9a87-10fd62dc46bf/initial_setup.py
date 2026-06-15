"""
Initial Setup: Apply gradient background fill to title cell A1
Task ID: calc_gfl_094
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Summary ---
    ws = wb.active
    ws.title = 'Summary'

    # Merge A1:F1 for title
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Executive Business Summary Q4 2024'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='1F3864')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    # Plain white background - NO gradient, NO fill
    ws['A1'].fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 20

    # Row 2: Headers
    headers = ['Department', 'Manager', 'Q4 Revenue', 'Q4 Expenses', 'Margin %', 'Status']
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20

    # Data rows 3-30 (28 rows of realistic business data)
    data = [
        ['Engineering', 'Sarah Chen', 1245000, 892000, 28.4, 'On Track'],
        ['Marketing', 'Marcus Johnson', 856000, 723000, 15.5, 'Under Review'],
        ['Sales', 'Priya Patel', 2134000, 1156000, 45.8, 'Exceeding'],
        ['Human Resources', 'David Kim', 312000, 298000, 4.5, 'On Track'],
        ['Finance', 'Elena Rodriguez', 478000, 356000, 25.5, 'On Track'],
        ['Operations', 'James Wright', 923000, 845000, 8.5, 'At Risk'],
        ['Product', 'Lisa Tanaka', 1567000, 1123000, 28.3, 'On Track'],
        ['Legal', 'Michael Brown', 215000, 201000, 6.5, 'Under Review'],
        ['Customer Success', 'Aisha Mohammed', 734000, 512000, 30.2, 'Exceeding'],
        ['Research & Dev', 'Thomas Mueller', 1890000, 1678000, 11.2, 'On Track'],
        ['Data Science', 'Yuki Nakamura', 567000, 423000, 25.4, 'On Track'],
        ['IT Infrastructure', 'Robert Garcia', 445000, 412000, 7.4, 'At Risk'],
        ['Quality Assurance', 'Anna Kowalski', 334000, 289000, 13.5, 'On Track'],
        ['Supply Chain', 'Carlos Mendez', 1123000, 987000, 12.1, 'Under Review'],
        ['Business Dev', 'Sophie Laurent', 678000, 445000, 34.4, 'Exceeding'],
        ['Compliance', 'Raj Krishnan', 198000, 187000, 5.6, 'On Track'],
        ['Training', 'Michelle Lee', 256000, 234000, 8.6, 'On Track'],
        ['Procurement', 'Hans Weber', 389000, 356000, 8.5, 'At Risk'],
        ['Facilities', 'Grace Okonkwo', 178000, 167000, 6.2, 'On Track'],
        ['Public Relations', 'Daniel Foster', 412000, 345000, 16.3, 'On Track'],
        ['Security', 'Nadia Petrov', 523000, 489000, 6.5, 'Under Review'],
        ['Analytics', 'Kevin Chang', 634000, 478000, 24.6, 'On Track'],
        ['Design', 'Isabella Rossi', 298000, 245000, 17.8, 'Exceeding'],
        ['Partnerships', 'Omar Hassan', 756000, 534000, 29.4, 'On Track'],
        ['Innovation Lab', 'Emma Wilson', 445000, 412000, 7.4, 'On Track'],
        ['Customer Support', 'Alex Nguyen', 389000, 356000, 8.5, 'At Risk'],
        ['Internal Audit', 'Rachel Adams', 167000, 156000, 6.6, 'On Track'],
        ['Strategy', 'William Blake', 534000, 423000, 20.8, 'On Track'],
    ]

    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.font = Font(name='Calibri', size=10)
            if c == 3 or c == 4:
                cell.number_format = '$#,##0'
            elif c == 5:
                cell.number_format = '0.0"%"'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
