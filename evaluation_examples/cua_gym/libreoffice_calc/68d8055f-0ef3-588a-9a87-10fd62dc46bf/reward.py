"""
Reward Script: Apply gradient background fill to title cell A1 (light blue to white)
Task ID: calc_gfl_094
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): A1 has a GradientFill (not PatternFill or none)
  Component 2 (0.3): Gradient colors are light blue (#ADD8E6) -> white (#FFFFFF)
  Component 3 (0.2): Gradient type is linear
  Component 4 (0.1): A1 still contains correct title text and merge is preserved
"""

import os
import sys

# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state():
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        import time
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_094'


def color_distance(rgb1, rgb2):
    """Compute simple RGB distance between two ARGB hex strings (8-char)."""
    try:
        # Extract RGB portion (ignore alpha)
        r1, g1, b1 = int(rgb1[2:4], 16), int(rgb1[4:6], 16), int(rgb1[6:8], 16)
        r2, g2, b2 = int(rgb2[2:4], 16), int(rgb2[4:6], 16), int(rgb2[6:8], 16)
        return ((r1 - r2) ** 2 + (g1 - g2) ** 2 + (b1 - b2) ** 2) ** 0.5
    except Exception:
        return 999


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Summary sheet
    try:
        ws = wb['Summary']
    except KeyError:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    cell = ws['A1']

    # Component 1: A1 has a GradientFill (0.4 points)
    # This is the primary change: initial has PatternFill (solid white), golden has GradientFill
    try:
        fill_type = cell.fill.fill_type
        # GradientFill types are 'linear' or 'path'; PatternFill types are 'solid', 'none', etc.
        is_gradient = fill_type in ('linear', 'path')
        if is_gradient:
            print(f"PASS: Component 1 — A1 has gradient fill (type: {fill_type}) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — A1 fill_type is '{fill_type}', expected gradient (linear/path)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Gradient colors are light blue -> white (0.3 points)
    # Expected: stop[0] ~ FFADD8E6 (light blue), stop[1] ~ FFFFFFFF (white)
    # Allow some tolerance for color variation (agent might pick slightly different light blue)
    try:
        stops = cell.fill.stop
        if stops and len(stops) >= 2:
            color0 = stops[0].color.rgb  # should be light blue
            color1 = stops[1].color.rgb  # should be white

            # Check first stop is a light blue (close to ADD8E6)
            dist_blue = color_distance(color0, 'FFADD8E6')
            # Also accept other common light blues
            dist_blue_alt1 = color_distance(color0, 'FF87CEEB')  # sky blue
            dist_blue_alt2 = color_distance(color0, 'FF4FC3F7')  # light blue variant
            min_blue_dist = min(dist_blue, dist_blue_alt1, dist_blue_alt2)

            # Check second stop is white or near-white
            dist_white = color_distance(color1, 'FFFFFFFF')

            # Also allow reversed direction (white -> light blue)
            dist_blue_rev = color_distance(color1, 'FFADD8E6')
            dist_blue_rev_alt1 = color_distance(color1, 'FF87CEEB')
            dist_blue_rev_alt2 = color_distance(color1, 'FF4FC3F7')
            min_blue_rev = min(dist_blue_rev, dist_blue_rev_alt1, dist_blue_rev_alt2)
            dist_white_rev = color_distance(color0, 'FFFFFFFF')

            # Forward: light blue -> white
            forward_ok = min_blue_dist < 80 and dist_white < 30
            # Reverse: white -> light blue
            reverse_ok = min_blue_rev < 80 and dist_white_rev < 30

            if forward_ok or reverse_ok:
                direction = "forward (blue->white)" if forward_ok else "reverse (white->blue)"
                print(f"PASS: Component 2 — Gradient colors correct ({direction}), "
                      f"stop0={color0}, stop1={color1} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Gradient colors not light-blue/white. "
                      f"stop0={color0} (blue_dist={min_blue_dist:.1f}), "
                      f"stop1={color1} (white_dist={dist_white:.1f})")
        else:
            print(f"FAIL: Component 2 — Gradient has {len(stops) if stops else 0} stops, expected >=2")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Gradient type is linear (0.2 points)
    try:
        fill_type = cell.fill.fill_type
        if fill_type == 'linear':
            print(f"PASS: Component 3 — Gradient type is 'linear' (0.2 pts)")
            total_score += 0.2
        elif fill_type == 'path':
            # Path gradient is also a valid gradient, partial credit
            print(f"PARTIAL: Component 3 — Gradient type is 'path' (not linear), awarding 0.1 pts")
            total_score += 0.1
        else:
            print(f"FAIL: Component 3 — fill_type is '{fill_type}', expected 'linear'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: A1 still has correct title and merge (0.1 points)
    # This checks that the gradient was applied without breaking the cell content
    # The merge and text are preconditions BUT we only award points if the gradient
    # is also present (compound check: gradient applied AND content preserved)
    try:
        has_gradient = cell.fill.fill_type in ('linear', 'path')
        title_text = cell.value
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in str(r) for r in ws.merged_cells.ranges)
        has_title = (title_text and 'Executive Business Summary' in str(title_text))

        if has_gradient and has_title and has_merge:
            print(f"PASS: Component 4 — Title text and merge preserved with gradient (0.1 pts)")
            total_score += 0.1
        elif not has_gradient:
            print(f"FAIL: Component 4 — No gradient found, skipping compound check")
        else:
            issues = []
            if not has_title:
                issues.append(f"title text missing (found: {title_text!r})")
            if not has_merge:
                issues.append(f"merge range missing (ranges: {merged_ranges})")
            print(f"FAIL: Component 4 — {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
