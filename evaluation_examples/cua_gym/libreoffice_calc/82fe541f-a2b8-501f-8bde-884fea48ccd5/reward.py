"""
Reward Script: Build a headcount summary section with department counts and a pie chart.
Task ID: calc_gpm_021
Domain: libreoffice_calc
Scoring:
  Component 1: Variance formulas in D2:D8 (0.20)
  Component 2: Total row A10:E10 with SUM formulas and bold (0.20)
  Component 3: Row 10 formatting - double top border + light purple fill (0.10)
  Component 4: Pie chart exists with correct title (0.20)
  Component 5: Pie chart percentage labels (0.10)
  Component 6: Conditional formatting on D2:D8 (0.10)
  Component 7: Data bars on B2:B8 (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_021'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Variance formulas in D2:D8 (0.20 points)
    # Initial has D2:D8 = None; golden has =B-C formulas
    try:
        variance_count = 0
        for r in range(2, 9):
            val = ws.cell(row=r, column=4).value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                # Accept =B2-C2 pattern or equivalent
                expected = f"=B{r}-C{r}"
                if normalized == expected.upper():
                    variance_count += 1
        if variance_count == 7:
            print(f"PASS: Component 1 - All 7 variance formulas found in D2:D8 (0.20 pts)")
            total_score += 0.20
        elif variance_count >= 4:
            partial = round(0.20 * variance_count / 7, 2)
            print(f"PARTIAL: Component 1 - {variance_count}/7 variance formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {variance_count}/7 variance formulas found in D2:D8")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Total row A10:E10 with SUM formulas and bold (0.20 points)
    # Initial has row 10 empty; golden has Total label + SUM formulas + bold
    try:
        comp2_score = 0.0

        # Check A10 = 'Total' and bold
        a10_val = ws.cell(row=10, column=1).value
        a10_bold = ws.cell(row=10, column=1).font.bold
        if a10_val is not None and str(a10_val).strip().lower() == 'total' and a10_bold:
            comp2_score += 0.05

        # Check B10, C10, D10, E10 have SUM formulas
        sum_count = 0
        for c in range(2, 6):
            val = ws.cell(row=10, column=c).value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                col_letter = chr(64 + c)
                expected = f"=SUM({col_letter}2:{col_letter}8)"
                if normalized == expected.upper():
                    sum_count += 1
        if sum_count == 4:
            comp2_score += 0.10

        # Check all row 10 cells are bold
        bold_count = 0
        for c in range(1, 6):
            if ws.cell(row=10, column=c).font.bold:
                bold_count += 1
        if bold_count == 5:
            comp2_score += 0.05

        if comp2_score > 0:
            print(f"PASS: Component 2 - Total row checks passed ({comp2_score} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 - Total row missing or incomplete (A10={a10_val}, SUM count={sum_count}, bold count={bold_count})")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Row 10 formatting - double top border + light purple fill (0.10 points)
    # Initial has no formatting on row 10; golden has double top border and FFE6D5F5 fill
    try:
        comp3_score = 0.0

        # Check double top border on row 10
        double_border_count = 0
        for c in range(1, 6):
            border_top = ws.cell(row=10, column=c).border.top.style if ws.cell(row=10, column=c).border.top else None
            if border_top == 'double':
                double_border_count += 1
        if double_border_count >= 3:
            comp3_score += 0.05

        # Check light purple fill on row 10
        purple_fill_count = 0
        for c in range(1, 6):
            try:
                fill_rgb = ws.cell(row=10, column=c).fill.fgColor.rgb
                # Accept any light purple-ish fill (not the default 00000000)
                if fill_rgb and fill_rgb != '00000000':
                    purple_fill_count += 1
            except Exception:
                pass
        if purple_fill_count >= 3:
            comp3_score += 0.05

        if comp3_score > 0:
            print(f"PASS: Component 3 - Row 10 formatting ({comp3_score} pts, double borders={double_border_count}, fill cells={purple_fill_count})")
            total_score += comp3_score
        else:
            print(f"FAIL: Component 3 - Row 10 missing double border or purple fill (double borders={double_border_count}, fill cells={purple_fill_count})")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Pie chart exists with correct title (0.20 points)
    # Initial has no charts; golden has a PieChart titled 'Headcount by Department'
    try:
        charts = ws._charts
        if len(charts) >= 1:
            pie_found = False
            title_match = False
            for chart in charts:
                if type(chart).__name__ == 'PieChart':
                    pie_found = True
                    # Extract title text
                    title_text = None
                    try:
                        if chart.title and chart.title.tx and chart.title.tx.rich:
                            for p in chart.title.tx.rich.p:
                                for run in p.r:
                                    if run.t:
                                        title_text = run.t
                    except Exception:
                        pass
                    if title_text and 'headcount' in title_text.lower() and 'department' in title_text.lower():
                        title_match = True
                    break

            if pie_found and title_match:
                print(f"PASS: Component 4 - Pie chart with title 'Headcount by Department' found (0.20 pts)")
                total_score += 0.20
            elif pie_found:
                print(f"PARTIAL: Component 4 - Pie chart found but title mismatch: '{title_text}' (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 - No PieChart found among {len(charts)} chart(s)")
        else:
            print(f"FAIL: Component 4 - No charts found in worksheet")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Pie chart percentage labels (0.10 points)
    # Initial has no chart; golden has showPercent=True
    try:
        charts = ws._charts
        pct_labels_found = False
        for chart in charts:
            if type(chart).__name__ == 'PieChart':
                if chart.dataLabels and chart.dataLabels.showPercent:
                    pct_labels_found = True
                break
        if pct_labels_found:
            print(f"PASS: Component 5 - Pie chart has percentage labels (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 - Pie chart missing percentage labels")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Conditional formatting on D2:D8 (0.10 points)
    # Initial has no CF; golden has 3 cellIs rules on D2:D8
    try:
        cf_on_d = False
        cf_rule_count = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'D2' in cf_range or 'D2:D8' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        cf_rule_count += 1
                if cf_rule_count >= 2:
                    cf_on_d = True
        if cf_on_d:
            print(f"PASS: Component 6 - Conditional formatting on D2:D8 with {cf_rule_count} cellIs rules (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 - No conditional formatting cellIs rules on D column (found {cf_rule_count})")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Data bars on B2:B8 (0.10 points)
    # Initial has no data bars; golden has dataBar CF on B2:B8
    try:
        data_bar_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'B2' in cf_range or 'B2:B8' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'dataBar':
                        data_bar_found = True
                        break
        if data_bar_found:
            print(f"PASS: Component 7 - Data bars found on B2:B8 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 - No data bars found on B column")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
