"""
Initial Setup: Build a headcount summary section with department counts and a pie chart.
Task ID: calc_gpm_021
Domain: libreoffice_calc

Creates the initial pre-task spreadsheet with raw headcount data.
The agent must: add D column formulas (=B-C), conditional formatting on D,
data bars on B, totals row 10 with formatting, and a pie chart.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Headcount'

    # --- Header row styling ---
    dark_purple = PatternFill(start_color="FF4B0082", end_color="FF4B0082", fill_type="solid")
    white_font_bold = Font(bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    all_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Headers
    headers = ['Department', 'Headcount', 'Budget HC', 'Variance', 'Open Positions']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = dark_purple
        cell.font = white_font_bold
        cell.alignment = center_align
        cell.border = all_border

    # Data rows 2-8
    data = [
        ['Engineering', 45, 50, None, 5],
        ['Marketing', 22, 25, None, 3],
        ['Sales', 38, 40, None, 2],
        ['Finance', 15, 15, None, 0],
        ['HR', 12, 14, None, 2],
        ['Operations', 28, 30, None, 2],
        ['Product', 18, 20, None, 2],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = all_border

    # Set column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
