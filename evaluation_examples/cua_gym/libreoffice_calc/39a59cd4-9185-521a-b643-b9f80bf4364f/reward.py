"""
Reward Script: Compute CLV for each customer segment and highlight those >2 StDev above mean
Task ID: osworld_calc_computed_col_highlight_max_007
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: CLV formulas present in column E (rows 2-13)     — 0.5 pts
  Component 2: Cells exceeding Mean+2*STDEV are highlighted green — 0.3 pts
  Component 3: Non-exceeding cells do NOT have a green background  — 0.2 pts
  Total: 1.0
"""

import os
import statistics

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_computed_col_highlight_max_007'


def is_green_fill(cell):
    """
    Returns True if a cell has a visibly green background fill.
    Checks for known green color (FF00B050) and any broadly green ARGB hex.
    """
    try:
        rgb = cell.fill.fgColor.rgb
        if rgb is None or rgb == '00000000':
            return False
        # Remove the alpha prefix (first 2 chars) to get 6-char RGB
        rgb6 = rgb[-6:].upper()
        r_val = int(rgb6[0:2], 16)
        g_val = int(rgb6[2:4], 16)
        b_val = int(rgb6[4:6], 16)
        # Green means G channel is dominant and noticeably present
        return g_val > 100 and g_val > r_val and g_val > b_val
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active/first sheet (expect 'Customer Segments')
    try:
        ws = wb.active
        print(f"INFO: Sheet = '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify source data columns exist (A-D, rows 2-13)
    expected_rows = range(2, 14)  # rows 2 through 13 (12 data rows)
    try:
        missing_data = []
        for r in expected_rows:
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            if b is None or c is None or d is None:
                missing_data.append(r)
        if missing_data:
            print(f"PRECONDITION FAIL: Rows {missing_data} are missing B/C/D source data. Cannot score.")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"PRECONDITION ERROR: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Compute expected CLV values from B, C, D columns
    try:
        clv_computed = {}
        for r in expected_rows:
            b = ws.cell(row=r, column=2).value
            c = ws.cell(row=r, column=3).value
            d = ws.cell(row=r, column=4).value
            clv_computed[r] = float(b) * float(c) * float(d)

        clv_values = list(clv_computed.values())
        mean_clv = statistics.mean(clv_values)
        stdev_clv = statistics.stdev(clv_values)
        threshold = mean_clv + 2 * stdev_clv

        print(f"INFO: Mean CLV = {mean_clv:.2f}, StDev = {stdev_clv:.2f}")
        print(f"INFO: Highlight threshold (Mean+2*StDev) = {threshold:.2f}")

        rows_above = [r for r, v in clv_computed.items() if v > threshold]
        rows_below = [r for r, v in clv_computed.items() if v <= threshold]
        print(f"INFO: Rows above threshold (should be green): {rows_above}")
        print(f"INFO: Rows at/below threshold (should NOT be green): {rows_below}")
    except Exception as e:
        print(f"CRITICAL: Could not compute expected CLV/threshold: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: CLV formulas present in E2:E13 (0.5 points)
    # A formula means the cell contains a string that starts with '='
    # and should follow the pattern =B*C*D (multiplication of B, C, D columns)
    try:
        formula_count = 0
        formula_failures = []
        for r in expected_rows:
            cell_e = ws.cell(row=r, column=5)
            val = cell_e.value
            if isinstance(val, str) and val.startswith('='):
                # Verify it contains multiplication of B, C, D references for this row
                val_upper = val.upper().replace(' ', '')
                expected_pattern = f'B{r}*C{r}*D{r}'
                # Also accept other orderings and/or PRODUCT formula
                has_b = f'B{r}' in val_upper
                has_c = f'C{r}' in val_upper
                has_d = f'D{r}' in val_upper
                has_mult = '*' in val_upper or 'PRODUCT' in val_upper
                if has_b and has_c and has_d and has_mult:
                    formula_count += 1
                else:
                    formula_failures.append((r, val))
            else:
                formula_failures.append((r, val))

        if formula_count == 12:
            print(f"PASS: Component 1 — All 12 CLV formulas present in E2:E13 (0.5 pts)")
            total_score += 0.5
        elif formula_count >= 6:
            partial = round(0.5 * formula_count / 12, 2)
            print(f"PARTIAL: Component 1 — {formula_count}/12 CLV formulas found, partial credit ({partial} pts)")
            print(f"  Failed rows: {formula_failures}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/12 CLV formulas found in column E")
            print(f"  Failed rows: {formula_failures}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cells exceeding Mean+2*STDEV are highlighted green (0.3 points)
    # Rows above threshold should have a green background fill
    try:
        if not rows_above:
            # No rows above threshold — check that all cells have no green fill (edge case)
            print("INFO: No rows above threshold — component 2 trivially passes")
            total_score += 0.3
        else:
            green_correct = 0
            green_failures = []
            for r in rows_above:
                cell_e = ws.cell(row=r, column=5)
                if is_green_fill(cell_e):
                    green_correct += 1
                else:
                    rgb = cell_e.fill.fgColor.rgb if cell_e.fill else 'N/A'
                    green_failures.append((r, rgb))

            if green_correct == len(rows_above):
                print(f"PASS: Component 2 — All {len(rows_above)} above-threshold row(s) highlighted green (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Only {green_correct}/{len(rows_above)} above-threshold rows highlighted green")
                print(f"  Failed rows: {green_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Exactly the correct number of cells are highlighted green (0.2 points)
    # This is a compound check: (a) cells above threshold have green fill AND
    # (b) cells below threshold do NOT have green fill — meaning the total count
    # of green-highlighted cells equals exactly the count of above-threshold rows.
    # This ensures that in the initial state (where NO cells are highlighted), the
    # check fails because the expected count (e.g. 1) != actual count (0).
    try:
        actual_green_rows = []
        for r in expected_rows:
            cell_e = ws.cell(row=r, column=5)
            if is_green_fill(cell_e):
                actual_green_rows.append(r)

        expected_green_rows = sorted(rows_above)
        actual_green_rows_sorted = sorted(actual_green_rows)

        # Check for false positives (cells highlighted that shouldn't be)
        false_positives = [r for r in actual_green_rows if r not in rows_above]
        # Check for false negatives (cells that should be highlighted but aren't)
        false_negatives = [r for r in rows_above if r not in actual_green_rows]

        if actual_green_rows_sorted == expected_green_rows and len(false_positives) == 0:
            print(f"PASS: Component 3 — Exactly {len(expected_green_rows)} cell(s) highlighted green, "
                  f"matching expected rows {expected_green_rows} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Green highlight count/rows mismatch")
            print(f"  Expected green rows: {expected_green_rows}")
            print(f"  Actual green rows:   {actual_green_rows_sorted}")
            if false_positives:
                print(f"  False positives (should NOT be green): {false_positives}")
            if false_negatives:
                print(f"  False negatives (should BE green): {false_negatives}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
