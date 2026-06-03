"""
Initial Setup: Compute customer lifetime value (CLV) and highlight outliers
Task ID: osworld_calc_computed_col_highlight_max_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_computed_col_highlight_max_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Customer Segments ---
    ws = wb.active
    ws.title = 'Customer Segments'

    # Headers (A=Segment, B=Avg Purchase Value, C=Purchase Frequency, D=Customer Lifespan, E=CLV)
    headers = ['Segment', 'Avg Purchase Value', 'Purchase Frequency', 'Customer Lifespan', 'CLV']
    header_fill = openpyxl.styles.PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = header_fill

    # Realistic customer segment data
    # Columns: Segment | Avg Purchase Value ($) | Purchase Frequency (per year) | Customer Lifespan (years)
    # Column E (CLV) left EMPTY — agent must compute B*C*D and enter formulas there
    data = [
        ('Casual Shoppers',      48.00,  4,  1.5),
        ('New Visitors',         62.00,  3,  1.0),
        ('Budget Buyers',        38.00,  6,  2.0),
        ('Infrequent Users',     85.00,  5,  2.5),
        ('Regular Members',     115.00,  8,  3.0),
        ('Engaged Loyalists',   175.00, 12,  4.0),
        ('Mobile-First',        140.00, 10,  3.5),
        ('Referral Channel',    195.00, 11,  4.5),
        ('Mid-Tier Clients',    265.00,  9,  5.0),
        ('Premium Segment',     385.00, 14,  5.5),
        ('Enterprise Accounts', 620.00, 20,  7.0),
        ('Strategic Partners',  980.00, 26,  9.0),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='left' if c == 1 else 'center')
        # Column E deliberately left empty — no CLV value yet

    # Column widths for readability
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched libreoffice --calc with DISPLAY=:0')


create_initial()
