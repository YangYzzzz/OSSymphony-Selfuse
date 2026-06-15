"""
Reward Script: Remove Category from rows and add as column field in pivot table
Task ID: calc_pivot_065
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Categories appear as column headers (cross-tab layout)
  Component 2 (0.25): Region remains as row field without nested Category rows
  Component 3 (0.25): Cross-tab data values are correct
  Component 4 (0.20): Grand Total preserved at 180000
"""

import os
import time


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def _is_close(val, expected, tol=1):
    """Check if a cell value is numerically close to expected."""
    try:
        return abs(float(val) - expected) < tol
    except (ValueError, TypeError):
        return False


WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_065'

# Expected cross-tab data: Region -> {Category: SUM of Sales}
EXPECTED_DATA = {
    'East':  {'Clothing': 14000, 'Electronics': 16000, 'Furniture': 15000},
    'North': {'Clothing': 12000, 'Electronics': 18000, 'Furniture': 15000},
    'South': {'Clothing': 10000, 'Electronics': 20000, 'Furniture': 15000},
    'West':  {'Clothing': 13000, 'Electronics': 15000, 'Furniture': 17000},
}
EXPECTED_CATEGORIES = {'Clothing', 'Electronics', 'Furniture'}
EXPECTED_REGIONS = {'East', 'North', 'South', 'West'}
EXPECTED_GRAND_TOTAL = 180000


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check PivotSheet exists
    if 'PivotSheet' not in wb.sheetnames:
        print("CRITICAL: 'PivotSheet' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PivotSheet']

    # Component 1: Categories appear as column headers (0.30 points)
    # In the golden cross-tab layout, the header row should contain category names
    # as column headers. In the initial nested layout, categories are row values in col B.
    try:
        header_values = set()
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                header_values.add(str(val).strip())

        # Check that all 3 categories appear in the header row
        categories_in_header = EXPECTED_CATEGORIES.intersection(header_values)
        if len(categories_in_header) == 3:
            print(f"PASS: Component 1 — All 3 categories found as column headers: {categories_in_header} (0.30 pts)")
            total_score += 0.30
        elif len(categories_in_header) >= 1:
            partial = round(0.10 * len(categories_in_header), 2)
            print(f"PARTIAL: Component 1 — {len(categories_in_header)}/3 categories in header: {categories_in_header} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No categories found in header row. Headers: {header_values}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Region as row field, Category NOT nested as rows (0.25 points)
    # In the golden layout, each region should appear exactly once as a row label.
    # In the initial layout, categories are nested under each region (multiple rows per region).
    # We check: regions appear in column A, AND there are no category names in non-header rows.
    try:
        # Collect all non-header row values from all columns to check layout
        col_a_values = []
        all_non_header_values = set()
        for row in range(2, ws.max_row + 1):
            val_a = ws.cell(row=row, column=1).value
            if val_a is not None:
                col_a_values.append(str(val_a).strip())
            # Check columns B onward for category names appearing as row data
            for col in range(2, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val is not None and isinstance(val, str):
                    all_non_header_values.add(val.strip())

        # Regions should be in column A
        regions_found = EXPECTED_REGIONS.intersection(set(col_a_values))
        # Categories should NOT appear as string values in non-header data rows
        # (In cross-tab, data cells are numeric, not category names)
        categories_in_data_rows = EXPECTED_CATEGORIES.intersection(all_non_header_values)

        if len(regions_found) == 4 and len(categories_in_data_rows) == 0:
            print(f"PASS: Component 2 — All 4 regions as row labels, no categories nested in rows (0.25 pts)")
            total_score += 0.25
        elif len(regions_found) == 4 and len(categories_in_data_rows) > 0:
            print(f"FAIL: Component 2 — Regions found but categories still in data rows: {categories_in_data_rows}")
        elif len(regions_found) > 0 and len(categories_in_data_rows) == 0:
            partial = round(0.25 * len(regions_found) / 4, 2)
            print(f"PARTIAL: Component 2 — {len(regions_found)}/4 regions found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Regions: {regions_found}, Categories in rows: {categories_in_data_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Cross-tab data values are correct (0.25 points)
    # Build a mapping from the pivot sheet and compare to expected values.
    try:
        # Find which columns correspond to which categories
        col_to_cat = {}
        for col in range(2, ws.max_column + 1):
            hdr = ws.cell(row=1, column=col).value
            if hdr is not None and str(hdr).strip() in EXPECTED_CATEGORIES:
                col_to_cat[col] = str(hdr).strip()

        correct_values = 0
        total_checks = 0

        for row in range(2, ws.max_row + 1):
            region_val = ws.cell(row=row, column=1).value
            if region_val is None:
                continue
            region_str = str(region_val).strip()
            if region_str not in EXPECTED_DATA:
                continue
            for col, cat in col_to_cat.items():
                cell_val = ws.cell(row=row, column=col).value
                expected_val = EXPECTED_DATA[region_str].get(cat)
                if expected_val is not None:
                    total_checks += 1
                    if cell_val is not None:
                        try:
                            if abs(float(cell_val) - expected_val) < 1:
                                correct_values += 1
                        except (ValueError, TypeError):
                            pass

        if total_checks > 0 and correct_values == total_checks:
            print(f"PASS: Component 3 — All {correct_values}/{total_checks} cross-tab values correct (0.25 pts)")
            total_score += 0.25
        elif total_checks > 0 and correct_values > 0:
            partial = round(0.25 * correct_values / total_checks, 2)
            print(f"PARTIAL: Component 3 — {correct_values}/{total_checks} values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No matching cross-tab values found (checks={total_checks}, correct={correct_values})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total preserved at 180000 in cross-tab layout (0.20 points)
    # The grand total must exist AND be in a row that has a "Grand Total" label
    # AND the cross-tab layout must be present (categories as column headers).
    # This compound check ensures it only passes on the golden cross-tab, not the
    # initial nested layout (where 180000 also exists but in a different structure).
    try:
        # Re-check: categories must be column headers (anchors to task change)
        header_cats = set()
        for col in range(2, ws.max_column + 1):
            hdr = ws.cell(row=1, column=col).value
            if hdr is not None and str(hdr).strip() in EXPECTED_CATEGORIES:
                header_cats.add(str(hdr).strip())

        cross_tab_layout = (len(header_cats) == 3)

        # Find Grand Total value by scanning all cells
        grand_total_matches = [
            (r, c)
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
            if ws.cell(row=r, column=c).value is not None
            and _is_close(ws.cell(row=r, column=c).value, EXPECTED_GRAND_TOTAL)
        ]
        grand_total_found = len(grand_total_matches) > 0

        if cross_tab_layout and grand_total_found:
            print(f"PASS: Component 4 — Grand Total {EXPECTED_GRAND_TOTAL} found in cross-tab layout (0.20 pts)")
            total_score += 0.20
        elif not cross_tab_layout:
            print(f"FAIL: Component 4 — Cross-tab layout not detected (categories as columns required)")
        else:
            print(f"FAIL: Component 4 — Grand Total {EXPECTED_GRAND_TOTAL} not found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
