"""
Initial Setup: Create a spreadsheet with raw data and a pivot-table-style summary
where Region and Category are nested row fields with SUM of Sales.
Task ID: calc_pivot_065
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # =====================================================
    # Sheet 1: Data (raw source data)
    # =====================================================
    ws_data = wb.active
    ws_data.title = "Data"

    headers = ["Region", "Category", "Sales"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Raw data: 4 regions x 3 categories = 12 base rows, with some duplicates for realism
    # Totals must sum to 180000
    raw_data = [
        # North: Electronics=18000, Clothing=12000, Furniture=15000 => 45000
        ["North", "Electronics", 10000],
        ["North", "Electronics", 8000],
        ["North", "Clothing", 5000],
        ["North", "Clothing", 7000],
        ["North", "Furniture", 9000],
        ["North", "Furniture", 6000],
        # South: Electronics=20000, Clothing=10000, Furniture=15000 => 45000
        ["South", "Electronics", 12000],
        ["South", "Electronics", 8000],
        ["South", "Clothing", 4000],
        ["South", "Clothing", 6000],
        ["South", "Furniture", 8000],
        ["South", "Furniture", 7000],
        # East: Electronics=16000, Clothing=14000, Furniture=15000 => 45000
        ["East", "Electronics", 9000],
        ["East", "Electronics", 7000],
        ["East", "Clothing", 8000],
        ["East", "Clothing", 6000],
        ["East", "Furniture", 10000],
        ["East", "Furniture", 5000],
        # West: Electronics=15000, Clothing=13000, Furniture=17000 => 45000
        ["West", "Electronics", 9000],
        ["West", "Electronics", 6000],
        ["West", "Clothing", 7000],
        ["West", "Clothing", 6000],
        ["West", "Furniture", 10000],
        ["West", "Furniture", 7000],
    ]

    for r, row_data in enumerate(raw_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_data.cell(row=r, column=c, value=val)

    ws_data.column_dimensions["A"].width = 12
    ws_data.column_dimensions["B"].width = 14
    ws_data.column_dimensions["C"].width = 12

    # =====================================================
    # Sheet 2: PivotSheet (pivot table with nested rows)
    # Region and Category as nested ROW fields, SUM of Sales
    # =====================================================
    ws_pivot = wb.create_sheet("PivotSheet")

    # Aggregated data per region+category
    # North: Electronics=18000, Clothing=12000, Furniture=15000 => 45000
    # South: Electronics=20000, Clothing=10000, Furniture=15000 => 45000
    # East:  Electronics=16000, Clothing=14000, Furniture=15000 => 45000
    # West:  Electronics=15000, Clothing=13000, Furniture=17000 => 45000
    # Grand Total: 180000

    pivot_header_font = Font(bold=True, size=11)
    pivot_header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    region_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")
    subtotal_fill = PatternFill(start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid")
    grand_fill = PatternFill(start_color="FFFCE4D6", end_color="FFFCE4D6", fill_type="solid")
    bold_font = Font(bold=True, size=11)
    currency_fmt = '#,##0'

    # Headers (row 1)
    ws_pivot.cell(row=1, column=1, value="Region").font = pivot_header_font
    ws_pivot.cell(row=1, column=1).fill = pivot_header_fill
    ws_pivot.cell(row=1, column=2, value="Category").font = pivot_header_font
    ws_pivot.cell(row=1, column=2).fill = pivot_header_fill
    ws_pivot.cell(row=1, column=3, value="SUM of Sales").font = pivot_header_font
    ws_pivot.cell(row=1, column=3).fill = pivot_header_fill

    # Nested row layout: Region spans multiple rows, Category nested underneath
    pivot_rows = [
        # (Region, Category, Sales, is_subtotal, is_grand)
        ("East", "Clothing", 14000, False, False),
        ("East", "Electronics", 16000, False, False),
        ("East", "Furniture", 15000, False, False),
        ("", "East Total", 45000, True, False),
        ("North", "Clothing", 12000, False, False),
        ("North", "Electronics", 18000, False, False),
        ("North", "Furniture", 15000, False, False),
        ("", "North Total", 45000, True, False),
        ("South", "Clothing", 10000, False, False),
        ("South", "Electronics", 20000, False, False),
        ("South", "Furniture", 15000, False, False),
        ("", "South Total", 45000, True, False),
        ("West", "Clothing", 13000, False, False),
        ("West", "Electronics", 15000, False, False),
        ("West", "Furniture", 17000, False, False),
        ("", "West Total", 45000, True, False),
        ("", "Grand Total", 180000, False, True),
    ]

    for i, (region, category, sales, is_sub, is_grand) in enumerate(pivot_rows, 2):
        c1 = ws_pivot.cell(row=i, column=1, value=region if region else "")
        c2 = ws_pivot.cell(row=i, column=2, value=category)
        c3 = ws_pivot.cell(row=i, column=3, value=sales)
        c3.number_format = currency_fmt

        if region:
            c1.font = bold_font
            c1.fill = region_fill
        if is_sub:
            c2.font = bold_font
            c2.fill = subtotal_fill
            c3.font = bold_font
            c3.fill = subtotal_fill
        if is_grand:
            c2.font = bold_font
            c2.fill = grand_fill
            c3.font = bold_font
            c3.fill = grand_fill

    ws_pivot.column_dimensions["A"].width = 14
    ws_pivot.column_dimensions["B"].width = 16
    ws_pivot.column_dimensions["C"].width = 16

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
