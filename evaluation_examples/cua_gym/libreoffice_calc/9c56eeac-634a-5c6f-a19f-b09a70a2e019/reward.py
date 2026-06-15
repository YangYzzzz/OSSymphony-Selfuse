"""
Reward Script: Build a product pricing model with cost-plus markup formulas,
margin calculations, column D highlighting, sheet protection, and data validation.
Task ID: calc_fin_product_pricing_075
Domain: libreoffice_calc
Scoring:
  Component 1: Pricing formulas C-E (0.25 pts)
  Component 2: Margin formulas F-H (0.20 pts)
  Component 3: Number formats B-H (0.15 pts)
  Component 4: Column D highlighted + bold header (0.15 pts)
  Component 5: Row 1 bold (0.05 pts)
  Component 6: Sheet protection with B column locked (0.10 pts)
  Component 7: Data validation on B2:B30 (0.10 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fin_product_pricing_075'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Pricing' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Pricing' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pricing']

    # Component 1: Pricing formulas in C2:C30, D2:D30, E2:E30 (0.25 points)
    # Expected: =B<n>*(1+0.30), =B<n>*(1+0.40), =B<n>*(1+0.50)
    try:
        price_formulas_correct = 0
        price_formulas_total = 0
        for row in range(2, 31):
            for col, factor in [(3, '0.30'), (4, '0.40'), (5, '0.50')]:
                price_formulas_total += 1
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if isinstance(val, str):
                    # Accept formula variants like =B2*(1+0.30) or =B2*1.30
                    normalized = val.upper().replace(' ', '')
                    expected1 = f'=B{row}*(1+{factor})'.upper()
                    expected2 = f'=B{row}*(1+{factor.rstrip("0")})'.upper()
                    if normalized == expected1 or normalized == expected2:
                        price_formulas_correct += 1

        ratio = price_formulas_correct / price_formulas_total if price_formulas_total > 0 else 0
        if ratio >= 0.95:
            print(f"PASS: Component 1 — Pricing formulas: {price_formulas_correct}/{price_formulas_total} correct (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = round(0.25 * ratio, 3)
            print(f"PARTIAL: Component 1 — Pricing formulas: {price_formulas_correct}/{price_formulas_total} correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Pricing formulas: only {price_formulas_correct}/{price_formulas_total} correct (expected =B<n>*(1+0.3x) pattern)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Margin formulas in F2:F30, G2:G30, H2:H30 (0.20 points)
    # Expected: =(C<n>-B<n>)/C<n>, =(D<n>-B<n>)/D<n>, =(E<n>-B<n>)/E<n>
    try:
        margin_correct = 0
        margin_total = 0
        col_letter = {6: 'C', 7: 'D', 8: 'E'}
        for row in range(2, 31):
            for col, price_col in [(6, 'C'), (7, 'D'), (8, 'E')]:
                margin_total += 1
                cell = ws.cell(row=row, column=col)
                val = cell.value
                if isinstance(val, str):
                    normalized = val.upper().replace(' ', '')
                    expected = f'=({price_col}{row}-B{row})/{price_col}{row}'.upper()
                    if normalized == expected:
                        margin_correct += 1

        ratio = margin_correct / margin_total if margin_total > 0 else 0
        if ratio >= 0.95:
            print(f"PASS: Component 2 — Margin formulas: {margin_correct}/{margin_total} correct (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = round(0.20 * ratio, 3)
            print(f"PARTIAL: Component 2 — Margin formulas: {margin_correct}/{margin_total} correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Margin formulas: only {margin_correct}/{margin_total} correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Number formats (0.15 points)
    # B2:E30 should be currency ($#,##0.00), F2:H30 should be percentage (0.00%)
    try:
        currency_correct = 0
        pct_correct = 0
        currency_total = 0
        pct_total = 0

        for row in range(2, 31):
            # B-E: currency
            for col in range(2, 6):
                currency_total += 1
                cell = ws.cell(row=row, column=col)
                fmt = cell.number_format or ''
                if '$' in fmt and '0.00' in fmt:
                    currency_correct += 1
            # F-H: percentage
            for col in range(6, 9):
                pct_total += 1
                cell = ws.cell(row=row, column=col)
                fmt = cell.number_format or ''
                if '%' in fmt:
                    pct_correct += 1

        currency_ratio = currency_correct / currency_total if currency_total > 0 else 0
        pct_ratio = pct_correct / pct_total if pct_total > 0 else 0

        if currency_ratio >= 0.95 and pct_ratio >= 0.95:
            print(f"PASS: Component 3 — Number formats: currency {currency_correct}/{currency_total}, pct {pct_correct}/{pct_total} (0.15 pts)")
            total_score += 0.15
        elif currency_ratio >= 0.5 or pct_ratio >= 0.5:
            partial = round(0.15 * (currency_ratio + pct_ratio) / 2, 3)
            print(f"PARTIAL: Component 3 — Number formats: currency {currency_correct}/{currency_total}, pct {pct_correct}/{pct_total} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Number formats: currency {currency_correct}/{currency_total}, pct {pct_correct}/{pct_total}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Column D highlighted (light blue background on D2:D30) + D1 bold (0.15 points)
    # Light blue = FFADD8E6
    try:
        d_highlighted = 0
        d_total = 29  # rows 2-30

        for row in range(2, 31):
            cell = ws.cell(row=row, column=4)
            try:
                fill_color = cell.fill.fgColor.rgb
                fill_type = cell.fill.fill_type
                # Check for light blue: FFADD8E6 or ADD8E6 (with or without alpha prefix)
                if fill_type == 'solid' and (fill_color == 'FFADD8E6' or fill_color.upper().endswith('ADD8E6')):
                    d_highlighted += 1
            except Exception:
                pass

        d1_bold = ws['D1'].font.bold is True

        highlight_ratio = d_highlighted / d_total
        if highlight_ratio >= 0.95 and d1_bold:
            print(f"PASS: Component 4 — Column D highlighted: {d_highlighted}/{d_total} cells light blue, D1 bold={d1_bold} (0.15 pts)")
            total_score += 0.15
        elif highlight_ratio >= 0.5 or d1_bold:
            partial = round(0.15 * ((highlight_ratio + (1.0 if d1_bold else 0.0)) / 2), 3)
            print(f"PARTIAL: Component 4 — Column D: {d_highlighted}/{d_total} highlighted, D1 bold={d1_bold} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Column D not highlighted: {d_highlighted}/{d_total} cells, D1 bold={d1_bold}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Row 1 bold (0.05 points)
    # All header cells A1:H1 should be bold
    try:
        bold_count = 0
        for col in range(1, 9):
            if ws.cell(row=1, column=col).font.bold is True:
                bold_count += 1

        if bold_count >= 8:
            print(f"PASS: Component 5 — Row 1 bold: {bold_count}/8 cells bold (0.05 pts)")
            total_score += 0.05
        elif bold_count >= 4:
            partial = round(0.05 * bold_count / 8, 3)
            print(f"PARTIAL: Component 5 — Row 1 bold: {bold_count}/8 cells bold ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Row 1 bold: only {bold_count}/8 cells bold")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Sheet protection enabled + B column locked (0.10 points)
    try:
        sheet_protected = ws.protection.sheet is True

        # Check that B column cells are locked
        b_locked_count = 0
        for row in range(2, 31):
            cell = ws.cell(row=row, column=2)
            if cell.protection.locked is True:
                b_locked_count += 1

        # C-E columns should not be locked (or at least B is locked differentially)
        b_locked = b_locked_count >= 25  # majority of B column locked

        if sheet_protected and b_locked:
            print(f"PASS: Component 6 — Sheet protected={sheet_protected}, B column locked ({b_locked_count}/29 cells) (0.10 pts)")
            total_score += 0.10
        elif sheet_protected:
            print(f"PARTIAL: Component 6 — Sheet protected but B column not locked (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Sheet not protected (protection.sheet={sheet_protected})")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Data validation on B2:B30 — decimal > 0, error message 'Cost must be positive' (0.10 points)
    try:
        validations = ws.data_validations.dataValidation
        dv_score = 0.0

        for dv in validations:
            sqref_str = str(dv.sqref).upper()
            if 'B2' in sqref_str and 'B30' in sqref_str:
                # Check it's decimal/whole type with greaterThan 0
                is_positive_check = (
                    dv.type in ('decimal', 'whole') and
                    dv.operator == 'greaterThan' and
                    str(dv.formula1) == '0'
                )
                has_error_msg = (
                    dv.error is not None and
                    'positive' in dv.error.lower()
                )
                if is_positive_check and has_error_msg:
                    dv_score = 0.10
                elif is_positive_check:
                    dv_score = max(dv_score, 0.05)
                else:
                    dv_score = max(dv_score, 0.02)
                break

        if dv_score >= 0.10:
            print(f"PASS: Component 7 — Data validation on B2:B30: decimal > 0 with error message (0.10 pts)")
            total_score += 0.10
        elif dv_score >= 0.05:
            print(f"PARTIAL: Component 7 — Data validation found on B2:B30 but missing error message or wrong type (0.05 pts)")
            total_score += dv_score
        else:
            print(f"FAIL: Component 7 — No data validation found on B2:B30")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
