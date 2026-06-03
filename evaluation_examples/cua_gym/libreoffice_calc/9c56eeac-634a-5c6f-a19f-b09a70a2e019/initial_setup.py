"""
Initial Setup: Product Pricing Model
Task ID: calc_fin_product_pricing_075
Domain: libreoffice_calc

Creates a Pricing sheet with product names and unit costs in columns A and B.
Columns C-H are empty (to be filled by the agent with pricing formulas).
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_product_pricing_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pricing'

    # --- Row 1: Headers (NOT bold - task will make them bold) ---
    headers = ['Product', 'Unit Cost', 'Price @30%', 'Price @40%', 'Price @50%',
               'Margin @30%', 'Margin @40%', 'Margin @50%']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Rows 2-30: Product data (realistic names and costs) ---
    # Only columns A (Product) and B (Unit Cost) are populated
    products = [
        ('Wireless Bluetooth Headphones', 42.50),
        ('Ergonomic Office Chair', 185.00),
        ('Stainless Steel Water Bottle', 8.75),
        ('Mechanical Keyboard USB', 67.20),
        ('Portable Power Bank 20000mAh', 23.40),
        ('Adjustable Standing Desk Mat', 31.60),
        ('Noise Cancelling Earbuds', 55.90),
        ('LED Desk Lamp with USB Port', 18.30),
        ('Bamboo Cutting Board Set', 12.80),
        ('Yoga Mat Non-Slip Premium', 22.15),
        ('Smart Watch Fitness Tracker', 78.50),
        ('Coffee Grinder Electric', 34.70),
        ('Insulated Lunch Box', 15.25),
        ('Laptop Stand Adjustable', 28.90),
        ('Ceramic Coffee Mug Set of 4', 19.60),
        ('Wireless Phone Charger Pad', 14.85),
        ('Reusable Shopping Bag Pack', 6.40),
        ('Electric Toothbrush Sonic', 45.20),
        ('Stainless Steel Cookware Set', 124.30),
        ('Portable Bluetooth Speaker', 38.75),
        ('USB-C Hub 7-in-1', 29.50),
        ('Air Purifier HEPA Filter', 89.00),
        ('Memory Foam Pillow', 27.80),
        ('Resistance Band Set', 11.40),
        ('Digital Kitchen Scale', 16.90),
        ('Wall Calendar 2025 Large', 7.20),
        ('Acrylic Paint Set 24 Colors', 21.35),
        ('Microfiber Towel Pack of 6', 13.60),
        ('Collapsible Silicone Straws', 5.85),
    ]

    for r, (product, cost) in enumerate(products, 2):
        ws.cell(row=r, column=1, value=product)
        ws.cell(row=r, column=2, value=cost)
        # Columns C-H are intentionally left empty

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Pricing')
    print(f'  Rows: 1 header + 29 data rows')
    print(f'  Columns A-B populated, C-H empty (ready for agent)')


create_initial()
