"""
Reward Script: Use FIXED() function to create formatted financial report strings in column G
Task ID: osworld_calc_text_format_number_006
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): Column G rows 2-13 are populated with formulas (not None/empty)
  - Component 2 (0.40): Each formula uses FIXED(Bx,2,FALSE) and FIXED(Cx,2,FALSE) for revenue and expenses
  - Component 3 (0.30): Each formula also includes FIXED(Dx,2,FALSE) for profit and matches required format pattern
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_text_format_number_006'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The golden state should have column G (rows 2-13) populated with formulas like:
    ="["&A2&"]: Rev=$"&FIXED(B2,2,FALSE)&" | Exp=$"&FIXED(C2,2,FALSE)&" | Profit=$"&FIXED(D2,2,FALSE)
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb["Financial_Data"]
    except KeyError:
        print("CRITICAL: Sheet 'Financial_Data' not found")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Column G rows 2-13 are all populated (not None/empty) (0.30 points)
    # This checks that the agent actually filled in column G for all 12 data rows.
    # Initial state: all G2:G13 are None/empty
    # Golden state: all G2:G13 contain formula strings
    # -----------------------------------------------------------------------
    try:
        populated_count = 0
        total_rows = 12  # rows 2..13

        for row_idx in range(2, 14):  # rows 2 to 13 inclusive
            cell_val = ws.cell(row=row_idx, column=7).value  # column G = 7
            if cell_val is not None and str(cell_val).strip() != "":
                populated_count += 1

        if populated_count == total_rows:
            print(f"PASS: Component 1 — all {total_rows} rows in G2:G13 are populated ({0.30} pts)")
            total_score += 0.30
        elif populated_count > 0:
            print(f"FAIL: Component 1 — only {populated_count}/{total_rows} rows in G2:G13 are populated")
        else:
            print(f"FAIL: Component 1 — column G rows 2-13 are all empty")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Each formula uses FIXED(Bx,2,FALSE) and FIXED(Cx,2,FALSE)
    # for Revenue and Expenses with 2 decimal places and thousands separators (0.40 points)
    # Initial state: G2:G13 are empty
    # Golden state: formulas contain FIXED(B<row>,2,FALSE) and FIXED(C<row>,2,FALSE)
    # -----------------------------------------------------------------------
    try:
        rev_exp_fixed_count = 0
        total_rows = 12  # rows 2..13

        for row_idx in range(2, 14):
            cell_val = ws.cell(row=row_idx, column=7).value
            if cell_val is None:
                continue
            formula_str = str(cell_val).upper().replace(" ", "")

            # Check for FIXED(B<row>,2,FALSE) - revenue
            rev_pattern = rf"FIXED\(B{row_idx},2,FALSE\)"
            # Check for FIXED(C<row>,2,FALSE) - expenses
            exp_pattern = rf"FIXED\(C{row_idx},2,FALSE\)"

            has_rev_fixed = bool(re.search(rev_pattern, formula_str))
            has_exp_fixed = bool(re.search(exp_pattern, formula_str))

            if has_rev_fixed and has_exp_fixed:
                rev_exp_fixed_count += 1

        if rev_exp_fixed_count == total_rows:
            print(f"PASS: Component 2 — all {total_rows} rows use FIXED(Bx,2,FALSE) and FIXED(Cx,2,FALSE) ({0.40} pts)")
            total_score += 0.40
        elif rev_exp_fixed_count > 0:
            print(f"FAIL: Component 2 — only {rev_exp_fixed_count}/{total_rows} rows have correct FIXED() for Revenue and Expenses")
        else:
            print(f"FAIL: Component 2 — no rows have correct FIXED() formulas for Revenue and Expenses")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Each formula includes FIXED(Dx,2,FALSE) for profit and
    # follows the pattern [Dept]: Rev=$... | Exp=$... | Profit=$... (0.30 points)
    # Initial state: G2:G13 are empty
    # Golden state: formulas include FIXED(D<row>,2,FALSE) and the text pattern with Rev=$, Exp=$, Profit=$
    # -----------------------------------------------------------------------
    try:
        profit_format_count = 0
        total_rows = 12  # rows 2..13

        for row_idx in range(2, 14):
            cell_val = ws.cell(row=row_idx, column=7).value
            if cell_val is None:
                continue
            formula_str = str(cell_val).upper().replace(" ", "")

            # Check for FIXED(D<row>,2,FALSE) - profit
            profit_pattern = rf"FIXED\(D{row_idx},2,FALSE\)"
            has_profit_fixed = bool(re.search(profit_pattern, formula_str))

            # Check for the required text format keywords: REV=$, EXP=$, PROFIT=$
            has_rev_label = "REV=$" in formula_str or 'REV=$"' in formula_str or '"REV=$"' in formula_str
            has_exp_label = "EXP=$" in formula_str or 'EXP=$"' in formula_str or '"EXP=$"' in formula_str
            has_profit_label = "PROFIT=$" in formula_str or 'PROFIT=$"' in formula_str or '"PROFIT=$"' in formula_str

            # Also check for the bracket notation for department: ["&A<row>&"]
            # In the formula, it should contain concatenation with A<row>
            a_col_pattern = rf"A{row_idx}"
            has_dept_ref = bool(re.search(a_col_pattern, formula_str))

            if has_profit_fixed and has_rev_label and has_exp_label and has_profit_label and has_dept_ref:
                profit_format_count += 1

        if profit_format_count == total_rows:
            print(f"PASS: Component 3 — all {total_rows} rows have profit FIXED() and correct format labels ({0.30} pts)")
            total_score += 0.30
        elif profit_format_count > 0:
            print(f"FAIL: Component 3 — only {profit_format_count}/{total_rows} rows have full format with profit and labels")
        else:
            print(f"FAIL: Component 3 — no rows have profit FIXED() with correct format labels")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 2)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
