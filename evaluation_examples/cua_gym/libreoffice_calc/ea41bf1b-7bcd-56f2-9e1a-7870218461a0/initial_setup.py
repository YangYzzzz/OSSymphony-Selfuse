"""
Initial Setup: Financial Report Spreadsheet - FIXED() Function Task
Task ID: osworld_calc_text_format_number_006
Domain: libreoffice_calc

Creates a spreadsheet with financial department data (Department, Revenue, Expenses, Profit)
where column G is empty. The agent must add FIXED() formulas to column G.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_text_format_number_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Financial_Data ---
    ws = wb.active
    ws.title = 'Financial_Data'

    # Header styling helpers
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=False)
    thin = Side(style='thin', color='AAAAAA')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column headers
    headers = ['Department', 'Revenue', 'Expenses', 'Profit', '', '', 'Summary']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Realistic financial department data (10 rows)
    # Columns: Department, Revenue, Expenses, Profit
    data = [
        ('Engineering',     2845320.75,  1923450.20,  921870.55),
        ('Marketing',       1534620.00,   987340.80,  547279.20),
        ('Sales',           4127890.50,  2341560.30, 1786330.20),
        ('Human Resources',  398750.00,   312890.45,   85859.55),
        ('Finance',          612340.80,   401230.60,  211110.20),
        ('Operations',      3298450.25,  2156780.90, 1141669.35),
        ('Research & Dev',  1876540.60,  1432890.75,  443649.85),
        ('Customer Support',  724680.40,   598340.20,  126340.20),
        ('Legal',            289430.00,   245670.90,   43759.10),
        ('Product Mgmt',    1045230.75,   723450.30,  321780.45),
        ('IT Infrastructure', 876540.00,  698230.50,  178309.50),
        ('Business Dev',    2134560.80,  1345670.40,  788890.40),
    ]

    data_font = Font(name='Calibri', size=11)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    money_format = '#,##0.00'

    for r, (dept, rev, exp, profit) in enumerate(data, 2):
        # Department
        cell_dept = ws.cell(row=r, column=1, value=dept)
        cell_dept.font = data_font
        cell_dept.border = data_border
        cell_dept.alignment = Alignment(horizontal='left')

        # Revenue
        cell_rev = ws.cell(row=r, column=2, value=rev)
        cell_rev.font = data_font
        cell_rev.border = data_border
        cell_rev.number_format = money_format
        cell_rev.alignment = Alignment(horizontal='right')

        # Expenses
        cell_exp = ws.cell(row=r, column=3, value=exp)
        cell_exp.font = data_font
        cell_exp.border = data_border
        cell_exp.number_format = money_format
        cell_exp.alignment = Alignment(horizontal='right')

        # Profit
        cell_profit = ws.cell(row=r, column=4, value=profit)
        cell_profit.font = data_font
        cell_profit.border = data_border
        cell_profit.number_format = money_format
        cell_profit.alignment = Alignment(horizontal='right')

        # Columns E and F are intentionally empty

        # Column G is intentionally EMPTY (agent must fill with FIXED() formulas)

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 5
    ws.column_dimensions['F'].width = 5
    ws.column_dimensions['G'].width = 65

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open with LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
