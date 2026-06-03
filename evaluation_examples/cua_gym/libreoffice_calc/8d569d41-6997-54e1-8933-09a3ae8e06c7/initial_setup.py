"""
Initial Setup: Warehouse receiving log with mixed quantity text entries
Task ID: calc_gen_data_cleanup_045
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_045'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ReceivingLog'

    # --- Headers ---
    headers = ['Receipt ID', 'Date', 'Supplier', 'Item', 'Quantity Text',
               'Qty Number', 'Unit', 'Std Unit', 'Review Flag']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Data rows (100 entries, A-E populated, F-I empty) ---
    suppliers = [
        'Apex Logistics', 'Summit Distributing', 'Harbor Freight Supply',
        'Northfield Industries', 'Coastal Wholesale', 'Prairie Distribution',
        'Delta Trading Co.', 'Keystone Suppliers', 'Meridian Goods',
        'Riverstone Imports'
    ]
    items = [
        'Steel Brackets', 'Rubber Gaskets', 'Copper Pipes', 'Plastic Bins',
        'Foam Packing', 'Metal Screws', 'PVC Couplings', 'Bolt Assortment',
        'Nylon Rope', 'Adhesive Tape', 'Silicone Sealant', 'Aluminum Foil',
        'Glass Bottles', 'Paper Bags', 'Cardboard Sheets', 'Wire Mesh',
        'Wooden Pallets', 'Plastic Film', 'Stainless Fasteners', 'Hose Clamps'
    ]
    units = ['boxes', 'pallets', 'units', 'cases', 'drums']
    # Weighted so pallets appear with some high quantities for meaningful task
    unit_weights = [30, 15, 30, 15, 10]

    # Fixed seed for reproducibility
    random.seed(42)

    dates = []
    import datetime
    base_date = datetime.date(2025, 1, 2)
    for i in range(100):
        d = base_date + datetime.timedelta(days=i // 2)
        dates.append(d.strftime('%Y-%m-%d'))

    for i in range(100):
        row = i + 2
        receipt_id = f'RCV-2025-{1000 + i:04d}'
        date_val = dates[i]
        supplier = suppliers[i % len(suppliers)]
        item = items[i % len(items)]

        # Pick unit with weights
        unit = random.choices(units, weights=unit_weights)[0]
        # Quantity: pallets tend to be smaller (1-20), others larger
        if unit == 'pallets':
            qty = random.randint(1, 20)
        elif unit == 'drums':
            qty = random.randint(1, 15)
        elif unit == 'boxes':
            qty = random.randint(5, 150)
        elif unit == 'cases':
            qty = random.randint(10, 200)
        else:  # units
            qty = random.randint(50, 500)

        qty_text = f'{qty} {unit}'

        ws.cell(row=row, column=1, value=receipt_id)
        ws.cell(row=row, column=2, value=date_val)
        ws.cell(row=row, column=3, value=supplier)
        ws.cell(row=row, column=4, value=item)
        ws.cell(row=row, column=5, value=qty_text)
        # Columns F, G, H, I left empty (task requires agent to fill them)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ReceivingLog with 100 data rows (rows 2-101)')
    print(f'  Columns A-E populated, F-I intentionally empty')

create_initial()
