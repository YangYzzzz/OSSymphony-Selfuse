"""
Reward Script: Warehouse Receiving Log Data Cleanup
Task ID: calc_gen_data_cleanup_045
Domain: libreoffice_calc
Scoring:
  - Component 1: Column F (Qty Number) has VALUE/LEFT/FIND extraction formulas for all 100 rows (0.25 pts)
  - Component 2: Column G (Unit) has MID/FIND extraction formulas for all 100 rows (0.25 pts)
  - Component 3: Column I (Review Flag) has IF/AND formula flagging pallets+qty>10 as REVIEW (0.25 pts)
  - Component 4: Column H (Std Unit) has dropdown data validation with standard units (0.15 pts)
  - Component 5: Conditional formatting with orange fill when I="REVIEW" (0.10 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_data_cleanup_045'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: ReceivingLog sheet must exist
    if 'ReceivingLog' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ReceivingLog' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ReceivingLog']

    # Component 1: Column F (Qty Number) — VALUE/LEFT/FIND formula for rows 2-101 (0.25 pts)
    # The formula extracts numeric part from E column: =VALUE(LEFT(E2,FIND(" ",E2)-1))
    # FAILS on initial (F cells all None), PASSES on golden (all 100 rows have formulas)
    try:
        f_formula_count = 0
        f_formula_correct = 0
        for row in range(2, 102):
            val = ws.cell(row=row, column=6).value  # Column F
            if val is not None:
                f_formula_count += 1
                val_str = str(val).upper()
                # Verify formula uses VALUE, LEFT, FIND to extract numeric portion from E column
                if ('VALUE' in val_str and 'LEFT' in val_str and 'FIND' in val_str
                        and f'E{row}' in str(val)):
                    f_formula_correct += 1

        if f_formula_count == 100 and f_formula_correct == 100:
            print("PASS: Component 1 — All 100 F column VALUE/LEFT/FIND extraction formulas present (0.25 pts)")
            total_score += 0.25
        elif f_formula_count == 100 and f_formula_correct > 50:
            partial = round(0.25 * (f_formula_correct / 100), 4)
            print(f"PARTIAL: Component 1 — {f_formula_correct}/100 F formulas correct ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — {f_formula_count}/100 F cells non-empty, {f_formula_correct} correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column G (Unit) — MID/FIND formula for rows 2-101 (0.25 pts)
    # The formula extracts text unit after the space: =MID(E2,FIND(" ",E2)+1,LEN(E2))
    # FAILS on initial (G cells all None), PASSES on golden (all 100 rows have formulas)
    try:
        g_formula_count = 0
        g_formula_correct = 0
        for row in range(2, 102):
            val = ws.cell(row=row, column=7).value  # Column G
            if val is not None:
                g_formula_count += 1
                val_str = str(val).upper()
                # Verify formula uses MID and FIND to extract text portion from E column
                if 'MID' in val_str and 'FIND' in val_str and f'E{row}' in str(val):
                    g_formula_correct += 1

        if g_formula_count == 100 and g_formula_correct == 100:
            print("PASS: Component 2 — All 100 G column MID/FIND extraction formulas present (0.25 pts)")
            total_score += 0.25
        elif g_formula_count == 100 and g_formula_correct > 50:
            partial = round(0.25 * (g_formula_correct / 100), 4)
            print(f"PARTIAL: Component 2 — {g_formula_correct}/100 G formulas correct ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 2 — {g_formula_count}/100 G cells non-empty, {g_formula_correct} correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column I (Review Flag) — IF/AND formula for rows 2-101 (0.25 pts)
    # Formula flags pallets with qty >10: =IF(AND(G2="pallets",F2>10),"REVIEW","")
    # FAILS on initial (I cells all None), PASSES on golden (all 100 rows have formulas)
    try:
        i_formula_count = 0
        i_formula_correct = 0
        for row in range(2, 102):
            val = ws.cell(row=row, column=9).value  # Column I
            if val is not None:
                i_formula_count += 1
                val_str = str(val).upper()
                # Verify formula uses IF/AND with pallets and REVIEW
                if ('IF' in val_str and 'AND' in val_str
                        and 'PALLETS' in val_str and 'REVIEW' in val_str
                        and f'G{row}' in str(val) and f'F{row}' in str(val)):
                    i_formula_correct += 1

        if i_formula_count == 100 and i_formula_correct == 100:
            print("PASS: Component 3 — All 100 I column IF/AND pallets>10 REVIEW formulas present (0.25 pts)")
            total_score += 0.25
        elif i_formula_count == 100 and i_formula_correct > 50:
            partial = round(0.25 * (i_formula_correct / 100), 4)
            print(f"PARTIAL: Component 3 — {i_formula_correct}/100 I formulas correct ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 3 — {i_formula_count}/100 I cells non-empty, {i_formula_correct} correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Column H (Std Unit) — Dropdown data validation with standard units (0.15 pts)
    # H2:H101 must have list validation: units, boxes, cases, pallets, drums
    # FAILS on initial (no data validation), PASSES on golden (H column validation present)
    try:
        validations = ws.data_validations.dataValidation
        h_validation_score = 0.0
        expected_units = {'units', 'boxes', 'cases', 'pallets', 'drums'}

        for dv in validations:
            sqref_str = str(dv.sqref)
            if 'H' in sqref_str and dv.type == 'list':
                # Validation on H column with list type found
                formula = str(dv.formula1).strip('"')
                dv_units = set(u.strip().lower() for u in formula.split(','))
                if expected_units.issubset(dv_units) or dv_units == expected_units:
                    h_validation_score = 0.15
                    print(f"PASS: Component 4 — H column dropdown with standard units found "
                          f"(units={sorted(dv_units)}) (0.15 pts)")
                elif len(dv_units.intersection(expected_units)) >= 3:
                    h_validation_score = 0.07
                    print(f"PARTIAL: Component 4 — H column has list validation but incomplete units "
                          f"(found={sorted(dv_units)}, expected={sorted(expected_units)}) (0.07 pts)")
                else:
                    h_validation_score = 0.0
                    print(f"FAIL: Component 4 — H column list validation has wrong units: {dv_units}")
                break

        if h_validation_score > 0:
            total_score += h_validation_score
        elif len(validations) == 0:
            print("FAIL: Component 4 — No data validation found on the sheet")
        else:
            print(f"FAIL: Component 4 — No list validation on H column ({len(validations)} validations on sheet)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting — orange fill when I column = "REVIEW" (0.10 pts)
    # FormulaRule applying to row range with $I="REVIEW" condition and orange/amber fill
    # FAILS on initial (no conditional formatting), PASSES on golden (CF rule present)
    try:
        cf_award = 0.0
        for cf_range in ws.conditional_formatting:
            for rule in ws.conditional_formatting[cf_range]:
                if rule.type in ('expression', 'formula') and rule.formula:
                    formula_upper = str(rule.formula).upper()
                    if 'REVIEW' in formula_upper:
                        # Review formula found — check for orange fill
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                                # Orange colors: FFFF9900, FFFFC000, FFFF8C00, FFFFA500
                                is_orange = (fill_color and (
                                    'FF9900' in fill_color.upper()
                                    or 'FFC000' in fill_color.upper()
                                    or 'FF8C00' in fill_color.upper()
                                    or 'FFA500' in fill_color.upper()
                                ))
                                if is_orange:
                                    cf_award = 0.10
                                    print(f"PASS: Component 5 — CF orange fill ({fill_color}) for REVIEW rows "
                                          f"found on range {cf_range} (0.10 pts)")
                                else:
                                    cf_award = 0.05
                                    print(f"PARTIAL: Component 5 — CF with REVIEW formula found but color "
                                          f"{fill_color} is not orange (0.05 pts)")
                            except Exception:
                                cf_award = 0.05
                                print("PARTIAL: Component 5 — CF with REVIEW formula found, fill color unreadable")
                        else:
                            cf_award = 0.05
                            print("PARTIAL: Component 5 — CF with REVIEW formula found but no fill defined")
                if cf_award > 0:
                    break
            if cf_award > 0:
                break

        if cf_award > 0:
            total_score += cf_award
        else:
            print("FAIL: Component 5 — No conditional formatting with REVIEW formula found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
