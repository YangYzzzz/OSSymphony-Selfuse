"""
Reward Script: Clinical Trial Patient Data Tracker
Task ID: calc_grs_061
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Change from Baseline formulas in column I
  Component 2 (0.20) - % Improvement formulas in column J
  Component 3 (0.20) - Summary Statistics formulas (Mean/Median/StdDev/Min/Max)
  Component 4 (0.15) - Treatment Arm average table with AVERAGEIF formulas
  Component 5 (0.15) - Conditional formatting (adverse events orange, withdrawals gray)
  Component 6 (0.10) - Line chart on Summary Statistics sheet
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_061'


def persist_app_state(domain: str):
    """Try to save any unsaved changes in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    if 'Patient Log' not in wb.sheetnames:
        print("CRITICAL: 'Patient Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'Summary Statistics' not in wb.sheetnames:
        print("CRITICAL: 'Summary Statistics' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_log = wb['Patient Log']
    ws_summary = wb['Summary Statistics']

    # Component 1: Change from Baseline formulas in column I (0.20 points)
    # Task requires: Change from Baseline = Week 12 Score - Baseline Score
    # Golden has: =H2-E2 pattern in I2:I31
    try:
        formula_count = 0
        valid_formulas = 0
        for r in range(2, 32):
            val = ws_log.cell(row=r, column=9).value  # Column I
            if val is not None:
                formula_count += 1
                val_str = str(val).upper().replace(" ", "")
                # Should reference H (Week 12) and E (Baseline) in a subtraction
                if val_str.startswith("=") and "H" in val_str and "E" in val_str:
                    valid_formulas += 1
        if valid_formulas >= 25:
            print(f"PASS: Component 1 — {valid_formulas}/30 Change from Baseline formulas found (0.20 pts)")
            total_score += 0.20
        elif valid_formulas >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 1 — {valid_formulas}/30 formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {valid_formulas} valid Change from Baseline formulas (need >=25)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: % Improvement formulas in column J (0.20 points)
    # Task requires: % Improvement = (Week 12 - Baseline) / Baseline * 100
    # Golden has: =(H2-E2)/E2*100 pattern in J2:J31
    try:
        formula_count = 0
        valid_formulas = 0
        for r in range(2, 32):
            val = ws_log.cell(row=r, column=10).value  # Column J
            if val is not None:
                formula_count += 1
                val_str = str(val).upper().replace(" ", "")
                # Should be a percentage formula referencing H and E with division
                if val_str.startswith("=") and "H" in val_str and "E" in val_str and "/" in val_str:
                    valid_formulas += 1
        if valid_formulas >= 25:
            print(f"PASS: Component 2 — {valid_formulas}/30 % Improvement formulas found (0.20 pts)")
            total_score += 0.20
        elif valid_formulas >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 2 — {valid_formulas}/30 formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {valid_formulas} valid % Improvement formulas (need >=25)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary Statistics formulas (0.20 points)
    # Golden has: Mean (AVERAGE), Median (MEDIAN), Std Dev (STDEV), Min (MIN), Max (MAX)
    # in rows 4-8, columns B-E referencing Patient Log time point columns
    try:
        expected_funcs = {
            4: 'AVERAGE',
            5: 'MEDIAN',
            6: 'STDEV',
            7: 'MIN',
            8: 'MAX',
        }
        stat_score = 0
        for row_num, func_name in expected_funcs.items():
            row_ok = 0
            for col in range(2, 6):  # B through E
                val = ws_summary.cell(row=row_num, column=col).value
                if val is not None:
                    val_str = str(val).upper().replace(" ", "")
                    if val_str.startswith("=") and func_name in val_str:
                        row_ok += 1
            if row_ok >= 3:  # At least 3 of 4 columns have the right formula
                stat_score += 1
        if stat_score >= 4:
            print(f"PASS: Component 3 — {stat_score}/5 summary statistic rows correct (0.20 pts)")
            total_score += 0.20
        elif stat_score >= 2:
            partial = 0.10
            print(f"PARTIAL: Component 3 — {stat_score}/5 stat rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {stat_score}/5 summary statistic rows correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Treatment Arm average table with AVERAGEIF (0.15 points)
    # Golden has rows 13-15 with AVERAGEIF formulas for Control, Treatment A, Treatment B
    try:
        treatment_arms_found = 0
        averageif_count = 0
        # Search for treatment arm labels and AVERAGEIF formulas in the Summary sheet
        for r in range(1, ws_summary.max_row + 1):
            cell_val = ws_summary.cell(row=r, column=1).value
            if cell_val and str(cell_val).strip() in ('Control', 'Treatment A', 'Treatment B'):
                treatment_arms_found += 1
                # Check if adjacent cells have AVERAGEIF formulas
                for col in range(2, 6):
                    formula = ws_summary.cell(row=r, column=col).value
                    if formula is not None:
                        formula_str = str(formula).upper().replace(" ", "")
                        if "AVERAGEIF" in formula_str:
                            averageif_count += 1
        if treatment_arms_found >= 3 and averageif_count >= 8:
            print(f"PASS: Component 4 — {treatment_arms_found} arms, {averageif_count} AVERAGEIF formulas (0.15 pts)")
            total_score += 0.15
        elif treatment_arms_found >= 2 and averageif_count >= 4:
            partial = 0.08
            print(f"PARTIAL: Component 4 — {treatment_arms_found} arms, {averageif_count} AVERAGEIF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — {treatment_arms_found} treatment arms found, {averageif_count} AVERAGEIF formulas")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting (0.15 points)
    # Golden has: expression rules for adverse events (K="Y" -> orange) and withdrawals (L="Y" -> gray)
    try:
        cf_rules_list = list(ws_log.conditional_formatting)
        # Count rules that reference K (adverse events) or L (withdrawal) columns
        has_adverse_rule = False
        has_withdrawal_rule = False
        for cf in cf_rules_list:
            for rule in cf.rules:
                formulas_str = ' '.join(str(f) for f in rule.formula) if rule.formula else ''
                formulas_upper = formulas_str.upper()
                # Check for adverse events rule (references column K with "Y")
                if '$K' in formulas_upper or 'K2' in formulas_upper.replace('$', ''):
                    if '"Y"' in formulas_str or "'Y'" in formulas_str:
                        has_adverse_rule = True
                # Check for withdrawal rule (references column L with "Y")
                if '$L' in formulas_upper or 'L2' in formulas_upper.replace('$', ''):
                    if '"Y"' in formulas_str or "'Y'" in formulas_str:
                        has_withdrawal_rule = True
        if has_adverse_rule and has_withdrawal_rule:
            print(f"PASS: Component 5 — Both conditional formatting rules found (0.15 pts)")
            total_score += 0.15
        elif has_adverse_rule or has_withdrawal_rule:
            partial = 0.08
            found = "adverse events" if has_adverse_rule else "withdrawal"
            print(f"PARTIAL: Component 5 — Only {found} rule found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules for adverse events or withdrawal")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Line chart on Summary Statistics sheet (0.10 points)
    # Golden has: 1 LineChart with title containing "Treatment Arm" or similar, with 3+ series
    try:
        charts = ws_summary._charts
        line_chart_found = False
        if len(charts) > 0:
            for chart in charts:
                chart_type = type(chart).__name__
                if 'Line' in chart_type:
                    line_chart_found = True
                    break
            if not line_chart_found:
                # Accept any chart type that has multiple series (could be rendered as line)
                for chart in charts:
                    if len(chart.series) >= 3:
                        line_chart_found = True
                        break
        if line_chart_found:
            print(f"PASS: Component 6 — Line chart found on Summary Statistics (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — No line chart found (charts: {len(charts)})")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
