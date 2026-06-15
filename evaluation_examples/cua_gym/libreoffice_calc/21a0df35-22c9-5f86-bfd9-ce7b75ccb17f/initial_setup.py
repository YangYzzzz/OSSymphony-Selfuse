"""
Initial Setup: Clinical trial patient data tracker
Task ID: calc_grs_061
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # =============================================
    # Sheet 1: Patient Log
    # =============================================
    ws = wb.active
    ws.title = "Patient Log"

    headers = [
        "Patient ID", "Enrollment Date", "Age Group", "Treatment Arm",
        "Baseline Score", "Week 4 Score", "Week 8 Score", "Week 12 Score",
        "Change from Baseline", "% Improvement", "Adverse Events", "Withdrawal"
    ]

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_widths = {
        'A': 12, 'B': 16, 'C': 12, 'D': 16,
        'E': 15, 'F': 14, 'G': 14, 'H': 16,
        'I': 20, 'J': 16, 'K': 16, 'L': 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # 30 patients: 10 Control, 10 Treatment A, 10 Treatment B
    # Realistic de-identified data
    patient_data = [
        # Control group (P001-P010)
        ["P001", "2025-01-08", "31-45", "Control", 72.3, 70.1, 68.5, 67.2, None, None, "N", "N"],
        ["P002", "2025-01-10", "46-60", "Control", 68.7, 67.2, 66.8, 65.9, None, None, "N", "N"],
        ["P003", "2025-01-12", "18-30", "Control", 75.1, 73.8, 72.4, 71.0, None, None, "N", "N"],
        ["P004", "2025-01-15", "61+", "Control", 64.5, 63.9, 62.1, 61.8, None, None, "Y", "N"],
        ["P005", "2025-01-18", "31-45", "Control", 70.2, 69.5, 68.3, 67.1, None, None, "N", "N"],
        ["P006", "2025-01-20", "46-60", "Control", 66.8, 65.4, 64.9, 63.7, None, None, "N", "Y"],
        ["P007", "2025-01-22", "18-30", "Control", 77.4, 76.1, 75.3, 74.8, None, None, "N", "N"],
        ["P008", "2025-01-25", "31-45", "Control", 69.1, 68.3, 67.0, 66.2, None, None, "Y", "N"],
        ["P009", "2025-01-28", "61+", "Control", 62.9, 61.7, 60.8, 59.5, None, None, "N", "N"],
        ["P010", "2025-02-01", "46-60", "Control", 71.6, 70.2, 69.1, 68.3, None, None, "N", "N"],
        # Treatment A group (P011-P020)
        ["P011", "2025-01-09", "31-45", "Treatment A", 73.5, 68.2, 63.7, 59.1, None, None, "N", "N"],
        ["P012", "2025-01-11", "46-60", "Treatment A", 69.8, 64.5, 60.2, 56.3, None, None, "N", "N"],
        ["P013", "2025-01-13", "18-30", "Treatment A", 76.2, 71.0, 66.4, 61.8, None, None, "N", "N"],
        ["P014", "2025-01-16", "61+", "Treatment A", 65.1, 61.3, 57.8, 54.2, None, None, "Y", "N"],
        ["P015", "2025-01-19", "31-45", "Treatment A", 71.4, 66.8, 62.1, 58.5, None, None, "N", "N"],
        ["P016", "2025-01-21", "46-60", "Treatment A", 67.9, 63.1, 59.4, 55.7, None, None, "N", "N"],
        ["P017", "2025-01-23", "18-30", "Treatment A", 78.3, 73.5, 68.9, 64.1, None, None, "Y", "Y"],
        ["P018", "2025-01-26", "31-45", "Treatment A", 70.6, 65.8, 61.2, 57.4, None, None, "N", "N"],
        ["P019", "2025-01-29", "61+", "Treatment A", 63.4, 59.1, 55.6, 52.0, None, None, "N", "N"],
        ["P020", "2025-02-02", "46-60", "Treatment A", 72.1, 67.4, 63.0, 58.7, None, None, "N", "N"],
        # Treatment B group (P021-P030)
        ["P021", "2025-01-09", "18-30", "Treatment B", 74.8, 67.3, 61.2, 55.4, None, None, "N", "N"],
        ["P022", "2025-01-12", "46-60", "Treatment B", 68.3, 62.1, 56.8, 51.9, None, None, "N", "N"],
        ["P023", "2025-01-14", "31-45", "Treatment B", 75.9, 68.7, 62.4, 56.8, None, None, "Y", "N"],
        ["P024", "2025-01-17", "61+", "Treatment B", 63.7, 58.2, 53.1, 48.5, None, None, "N", "N"],
        ["P025", "2025-01-20", "18-30", "Treatment B", 77.1, 70.4, 64.0, 58.2, None, None, "N", "N"],
        ["P026", "2025-01-22", "46-60", "Treatment B", 66.5, 60.3, 55.1, 50.4, None, None, "Y", "Y"],
        ["P027", "2025-01-24", "31-45", "Treatment B", 73.2, 66.5, 60.8, 55.1, None, None, "N", "N"],
        ["P028", "2025-01-27", "18-30", "Treatment B", 79.4, 72.1, 65.7, 59.8, None, None, "N", "N"],
        ["P029", "2025-01-30", "61+", "Treatment B", 61.8, 56.9, 52.3, 47.6, None, None, "N", "N"],
        ["P030", "2025-02-03", "31-45", "Treatment B", 71.5, 65.0, 59.2, 54.0, None, None, "N", "N"],
    ]

    data_align = Alignment(horizontal="center", vertical="center")
    for r, row_data in enumerate(patient_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = data_align
            cell.border = thin_border
            # Number format for scores
            if c in (5, 6, 7, 8) and val is not None:
                cell.number_format = '0.0'

    # Data Validation: Age Group dropdown (column C)
    dv_age = DataValidation(
        type="list",
        formula1='"18-30,31-45,46-60,61+"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_age.error = "Please select a valid age group"
    dv_age.errorTitle = "Invalid Age Group"
    dv_age.add("C2:C100")
    ws.add_data_validation(dv_age)

    # Data Validation: Treatment Arm dropdown (column D)
    dv_arm = DataValidation(
        type="list",
        formula1='"Control,Treatment A,Treatment B"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_arm.error = "Please select a valid treatment arm"
    dv_arm.errorTitle = "Invalid Treatment Arm"
    dv_arm.add("D2:D100")
    ws.add_data_validation(dv_arm)

    # Data Validation: Adverse Events Y/N dropdown (column K)
    dv_ae = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_ae.add("K2:K100")
    ws.add_data_validation(dv_ae)

    # Data Validation: Withdrawal Y/N dropdown (column L)
    dv_wd = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_wd.add("L2:L100")
    ws.add_data_validation(dv_wd)

    # NOTE: Change from Baseline (col I) and % Improvement (col J) are LEFT EMPTY
    # The task requires the agent to add formulas there.
    # NOTE: NO conditional formatting applied - task requires the agent to do it.
    # NOTE: NO chart - task requires the agent to create it.

    # =============================================
    # Sheet 2: Summary Statistics (empty template)
    # =============================================
    ws2 = wb.create_sheet("Summary Statistics")

    # Headers for summary table - structure only, NO formulas
    ws2["A1"] = "Summary Statistics by Time Point"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws2.merge_cells("A1:F1")

    # Sub-headers
    stat_headers = ["Statistic", "Baseline", "Week 4", "Week 8", "Week 12"]
    for col, h in enumerate(stat_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Row labels only - NO formulas (agent must fill these)
    stat_labels = ["Mean", "Median", "Std Dev", "Min", "Max"]
    for r, label in enumerate(stat_labels, 4):
        cell = ws2.cell(row=r, column=1, value=label)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.border = thin_border
        # Empty cells for the formulas
        for c in range(2, 6):
            ws2.cell(row=r, column=c).border = thin_border

    # Column widths
    ws2.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E']:
        ws2.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
