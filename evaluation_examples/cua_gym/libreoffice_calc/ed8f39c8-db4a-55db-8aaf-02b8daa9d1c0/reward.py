"""
Reward Script: Case-insensitive lookup using LOWER on both sides
Task ID: calc_fma_vlookup_case_insensitive_057
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): At least one formula in B2:B11 uses LOWER for case-insensitive lookup
  Component 2 (0.3): All 10 cells B2:B11 contain LOWER-based lookup formulas
  Component 3 (0.3): Formulas reference correct catalog range (D2:D11) and price range (E2:E11)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_vlookup_case_insensitive_057'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'CatalogLookup' must exist
    if 'CatalogLookup' not in wb.sheetnames:
        print("FAIL: Sheet 'CatalogLookup' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CatalogLookup']

    # Helper: check if a value looks like a LOWER-based case-insensitive lookup formula
    def is_lower_lookup_formula(val):
        """Returns True if val is a formula string containing LOWER and a lookup function."""
        if not isinstance(val, str):
            return False
        val_upper = val.upper().replace(' ', '')
        # Must be a formula
        if not val_upper.startswith('='):
            return False
        # Must use LOWER (for case-insensitive matching)
        if 'LOWER' not in val_upper:
            return False
        # Must use some form of lookup: VLOOKUP, MATCH, INDEX, HLOOKUP
        has_lookup = any(fn in val_upper for fn in ['VLOOKUP', 'MATCH(', 'INDEX(', 'HLOOKUP'])
        return has_lookup

    # Component 1: At least one formula in B2:B11 uses LOWER for case-insensitive lookup (0.4 points)
    try:
        formulas_with_lower = []
        for row in range(2, 12):  # rows 2 to 11 inclusive
            cell_val = ws.cell(row=row, column=2).value
            if is_lower_lookup_formula(cell_val):
                formulas_with_lower.append(f"B{row}")

        if len(formulas_with_lower) >= 1:
            print(f"PASS: Component 1 — Found LOWER-based lookup formula(s) in: {formulas_with_lower[:3]}... ({len(formulas_with_lower)} total) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — No LOWER-based lookup formulas found in B2:B11")
            # Print actual values for debugging
            for row in range(2, 12):
                print(f"  B{row} = {repr(ws.cell(row=row, column=2).value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 10 cells B2:B11 contain LOWER-based lookup formulas (0.3 points)
    try:
        missing = []
        for row in range(2, 12):
            cell_val = ws.cell(row=row, column=2).value
            if not is_lower_lookup_formula(cell_val):
                missing.append(f"B{row}")

        if len(missing) == 0:
            print(f"PASS: Component 2 — All 10 cells B2:B11 have LOWER-based lookup formulas (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Missing or invalid formulas in: {missing}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas reference correct data ranges D2:D11 (catalog) and E2:E11 (prices) (0.3 points)
    try:
        correct_range_count = 0
        for row in range(2, 12):
            cell_val = ws.cell(row=row, column=2).value
            if isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(' ', '')
                # Check for references to D column (catalog names) and E column (prices)
                # Accept both absolute ($D$2:$D$11) and relative (D2:D11) references
                has_d_ref = bool(re.search(r'\$?D\$?[0-9]', val_upper))
                has_e_ref = bool(re.search(r'\$?E\$?[0-9]', val_upper))
                if has_d_ref and has_e_ref:
                    correct_range_count += 1

        if correct_range_count == 10:
            print(f"PASS: Component 3 — All 10 formulas reference both D (catalog) and E (price) columns (0.3 pts)")
            total_score += 0.3
        elif correct_range_count >= 1:
            print(f"FAIL: Component 3 — Only {correct_range_count}/10 formulas reference correct D and E columns")
        else:
            print(f"FAIL: Component 3 — No formulas reference both D and E columns correctly")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
