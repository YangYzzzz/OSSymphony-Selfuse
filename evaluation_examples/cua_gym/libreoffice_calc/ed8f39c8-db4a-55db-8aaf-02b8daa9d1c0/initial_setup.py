"""
Initial Setup: Case-insensitive VLOOKUP using LOWER on both sides
Task ID: calc_fma_vlookup_case_insensitive_057
Domain: libreoffice_calc

Creates a spreadsheet with a CatalogLookup sheet containing:
- Column A: Search terms with mixed case (APPLE, banana, Cherry, etc.)
- Column B: Empty Price column (to be filled by the agent)
- Column D: Catalog fruit names in title case
- Column E: Corresponding prices
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_vlookup_case_insensitive_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CatalogLookup ---
    ws = wb.active
    ws.title = 'CatalogLookup'

    # Headers in row 1
    ws['A1'] = 'Search Term'
    ws['B1'] = 'Price'
    ws['D1'] = 'Fruit'
    ws['E1'] = 'Catalog Price'

    # Style headers with bold font
    for coord in ['A1', 'B1', 'D1', 'E1']:
        ws[coord].font = Font(bold=True)
        ws[coord].alignment = Alignment(horizontal='center')

    # Column A: Search terms with mixed case (intentionally varied capitalization)
    search_terms = [
        'APPLE',       # all caps
        'banana',      # all lower
        'Cherry',      # title case
        'DURIAN',      # all caps
        'elderberry',  # all lower
        'FIG',         # all caps
        'grape',       # all lower
        'HONEYDEW',    # all caps
        'jackfruit',   # all lower
        'KIWI',        # all caps
    ]

    # Column D: Catalog fruit names in title case
    catalog_fruits = [
        'Apple',
        'Banana',
        'Cherry',
        'Durian',
        'Elderberry',
        'Fig',
        'Grape',
        'Honeydew',
        'Jackfruit',
        'Kiwi',
    ]

    # Column E: Prices corresponding to catalog entries
    catalog_prices = [1.20, 0.50, 2.00, 3.50, 4.00, 1.80, 0.90, 1.50, 5.00, 2.20]

    # Write data rows 2-11
    for i, (term, fruit, price) in enumerate(zip(search_terms, catalog_fruits, catalog_prices), start=2):
        ws.cell(row=i, column=1, value=term)      # Column A: search term (mixed case)
        # Column B (column 2): intentionally left empty - agent must fill this
        ws.cell(row=i, column=4, value=fruit)     # Column D: catalog fruit
        ws.cell(row=i, column=5, value=price)     # Column E: catalog price

    # Column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 4   # spacer column
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: CatalogLookup')
    print(f'  Column A (rows 2-11): Mixed-case search terms')
    print(f'  Column B (rows 2-11): EMPTY (to be filled by agent)')
    print(f'  Column D (rows 2-11): Title-case catalog fruits')
    print(f'  Column E (rows 2-11): Catalog prices')


create_initial()
