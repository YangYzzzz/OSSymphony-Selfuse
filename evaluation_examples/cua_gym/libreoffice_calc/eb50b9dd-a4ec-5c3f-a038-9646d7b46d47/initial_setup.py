"""
Initial Setup: Create a Status Report spreadsheet with no print scaling applied
Task ID: calc_adv_print_scale_pct_023
Domain: libreoffice_calc

Creates a realistic project status report with data in rows 1-60, columns A-H,
with A4 portrait paper settings and default (100%) print scale.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_print_scale_pct_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Status Report ---
    ws = wb.active
    ws.title = 'Status Report'

    # Header row styling
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column headers (A-H)
    headers = [
        'Task ID', 'Project Name', 'Assigned To', 'Department',
        'Status', 'Priority', 'Due Date', 'Completion %'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Realistic project status report data (rows 2-60 = 59 data rows)
    data = [
        ('T-001', 'Website Redesign', 'Sarah Chen', 'Marketing', 'In Progress', 'High', '2025-04-15', 65),
        ('T-002', 'ERP System Upgrade', 'Marcus Johnson', 'IT', 'Planning', 'Critical', '2025-05-01', 20),
        ('T-003', 'Annual Audit Prep', 'Linda Hoffman', 'Finance', 'Complete', 'High', '2025-03-31', 100),
        ('T-004', 'Staff Training Program', 'David Park', 'HR', 'In Progress', 'Medium', '2025-04-30', 45),
        ('T-005', 'Supply Chain Review', 'Emma Williams', 'Operations', 'Delayed', 'High', '2025-03-20', 30),
        ('T-006', 'Product Launch Campaign', 'Nathan Torres', 'Marketing', 'In Progress', 'Critical', '2025-04-10', 55),
        ('T-007', 'Office Relocation Plan', 'Sandra Kim', 'Admin', 'Planning', 'Medium', '2025-06-01', 10),
        ('T-008', 'Customer Portal v2', 'James Nguyen', 'IT', 'In Progress', 'High', '2025-05-15', 40),
        ('T-009', 'Budget Reforecast Q2', 'Carol Martinez', 'Finance', 'Complete', 'Medium', '2025-03-25', 100),
        ('T-010', 'Vendor Contract Renewal', 'Robert Lewis', 'Procurement', 'In Review', 'High', '2025-04-05', 80),
        ('T-011', 'Security Policy Update', 'Angela Davis', 'IT', 'In Progress', 'Critical', '2025-04-20', 60),
        ('T-012', 'Warehouse Optimization', 'Brian Scott', 'Operations', 'Planning', 'Medium', '2025-05-30', 15),
        ('T-013', 'Mobile App Release', 'Priya Sharma', 'Engineering', 'In Progress', 'High', '2025-04-25', 70),
        ('T-014', 'Compliance Training', 'Kevin White', 'HR', 'Complete', 'Medium', '2025-03-15', 100),
        ('T-015', 'Data Center Migration', 'Megan Brown', 'IT', 'Delayed', 'Critical', '2025-04-08', 35),
        ('T-016', 'Social Media Strategy', 'Tyler Adams', 'Marketing', 'In Progress', 'Medium', '2025-05-10', 50),
        ('T-017', 'Payroll System Audit', 'Helen Clark', 'Finance', 'In Review', 'High', '2025-04-02', 90),
        ('T-018', 'New Hire Onboarding', 'Christopher Hall', 'HR', 'In Progress', 'Medium', '2025-04-18', 55),
        ('T-019', 'Client Reporting Dashboard', 'Natalie Young', 'Engineering', 'In Progress', 'High', '2025-05-05', 45),
        ('T-020', 'ISO 27001 Certification', 'Daniel Robinson', 'IT', 'Planning', 'Critical', '2025-07-01', 5),
        ('T-021', 'Fleet Management Review', 'Lisa Walker', 'Operations', 'Complete', 'Low', '2025-03-10', 100),
        ('T-022', 'Brand Identity Refresh', 'Alex Peterson', 'Marketing', 'In Progress', 'Medium', '2025-05-20', 30),
        ('T-023', 'Benefits Package Update', 'Olivia Turner', 'HR', 'In Review', 'Medium', '2025-04-12', 75),
        ('T-024', 'API Integration Project', 'Ryan Mitchell', 'Engineering', 'In Progress', 'High', '2025-04-28', 60),
        ('T-025', 'Sustainability Report', 'Jessica Moore', 'Admin', 'Planning', 'Low', '2025-06-15', 10),
        ('T-026', 'Customer Feedback Analysis', 'Benjamin Lee', 'Marketing', 'In Progress', 'Medium', '2025-04-22', 40),
        ('T-027', 'Software License Audit', 'Stephanie Taylor', 'IT', 'Complete', 'High', '2025-03-28', 100),
        ('T-028', 'Product Roadmap Review', 'Aaron Jackson', 'Engineering', 'In Progress', 'High', '2025-05-08', 50),
        ('T-029', 'Accounts Receivable Review', 'Michelle Harris', 'Finance', 'In Review', 'Medium', '2025-04-06', 85),
        ('T-030', 'Team Building Events Q2', 'Jonathan Martin', 'HR', 'Planning', 'Low', '2025-05-25', 20),
        ('T-031', 'Network Infrastructure Upgrade', 'Rebecca Garcia', 'IT', 'In Progress', 'Critical', '2025-05-12', 35),
        ('T-032', 'Retail Expansion Analysis', 'Charles Thompson', 'Operations', 'Planning', 'High', '2025-06-30', 10),
        ('T-033', 'Email Marketing Campaign', 'Amanda Wilson', 'Marketing', 'Complete', 'Medium', '2025-03-22', 100),
        ('T-034', 'Financial Controls Review', 'Patrick Anderson', 'Finance', 'In Progress', 'High', '2025-04-16', 65),
        ('T-035', 'Cloud Storage Migration', 'Rachel Thomas', 'IT', 'In Progress', 'Medium', '2025-05-18', 45),
        ('T-036', 'Employee Satisfaction Survey', 'Gregory Jackson', 'HR', 'Complete', 'Low', '2025-03-05', 100),
        ('T-037', 'Logistics Route Optimization', 'Victoria White', 'Operations', 'In Progress', 'Medium', '2025-04-24', 55),
        ('T-038', 'UX Research Study', 'Samuel Hill', 'Engineering', 'In Review', 'High', '2025-04-14', 90),
        ('T-039', 'Corporate Website Update', 'Kimberly Lopez', 'Marketing', 'In Progress', 'Medium', '2025-05-06', 35),
        ('T-040', 'Procurement Policy Revision', 'Timothy Green', 'Procurement', 'Complete', 'Medium', '2025-03-18', 100),
        ('T-041', 'CRM System Deployment', 'Christine Baker', 'IT', 'In Progress', 'Critical', '2025-05-22', 40),
        ('T-042', 'Quarterly Management Review', 'Lawrence Adams', 'Admin', 'In Progress', 'High', '2025-04-03', 70),
        ('T-043', 'Cost Reduction Initiative', 'Dorothy Nelson', 'Finance', 'Planning', 'High', '2025-06-01', 15),
        ('T-044', 'Product Quality Assessment', 'Harold Carter', 'Operations', 'In Review', 'Medium', '2025-04-09', 80),
        ('T-045', 'Technical Documentation', 'Barbara Mitchell', 'Engineering', 'In Progress', 'Low', '2025-05-14', 60),
        ('T-046', 'Digital Advertising Campaign', 'Frank Perez', 'Marketing', 'In Progress', 'Medium', '2025-04-26', 50),
        ('T-047', 'Benefits Administration Audit', 'Margaret Roberts', 'HR', 'Complete', 'Medium', '2025-03-12', 100),
        ('T-048', 'Server Hardware Refresh', 'Walter Turner', 'IT', 'Planning', 'High', '2025-05-28', 10),
        ('T-049', 'Supplier Performance Review', 'Betty Phillips', 'Procurement', 'In Review', 'Medium', '2025-04-11', 85),
        ('T-050', 'Office Equipment Inventory', 'Ernest Campbell', 'Admin', 'Complete', 'Low', '2025-03-08', 100),
        ('T-051', 'Customer Service Training', 'Frances Evans', 'HR', 'In Progress', 'Medium', '2025-04-29', 55),
        ('T-052', 'Machine Learning POC', 'Howard Collins', 'Engineering', 'In Progress', 'High', '2025-06-10', 25),
        ('T-053', 'Trade Show Preparation', 'Ruth Stewart', 'Marketing', 'In Progress', 'High', '2025-04-07', 75),
        ('T-054', 'Tax Filing Preparation', 'Roy Sanchez', 'Finance', 'Complete', 'Critical', '2025-03-30', 100),
        ('T-055', 'Production Line Review', 'Evelyn Morris', 'Operations', 'In Progress', 'Medium', '2025-05-02', 40),
        ('T-056', 'Disaster Recovery Testing', 'Raymond Rogers', 'IT', 'Delayed', 'Critical', '2025-04-04', 30),
        ('T-057', 'Content Management System', 'Lois Reed', 'Engineering', 'In Progress', 'Medium', '2025-05-16', 50),
        ('T-058', 'Partnership Agreement Review', 'Harold Cook', 'Admin', 'In Review', 'High', '2025-04-17', 80),
        ('T-059', 'Customer Retention Program', 'Marie Bell', 'Marketing', 'Planning', 'Medium', '2025-06-05', 10),
        ('T-060', 'Year-End Financial Close', 'Fred Murphy', 'Finance', 'In Progress', 'Critical', '2025-04-01', 60),
    ]

    # Fill data rows (row 2 to row 60)
    cell_align = Alignment(vertical='center')
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = cell_align
            cell.border = cell_border

    # Column widths
    col_widths = [8, 28, 20, 14, 12, 10, 12, 14]
    col_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    for letter, width in zip(col_letters, col_widths):
        ws.column_dimensions[letter].width = width

    # Row 1 header height
    ws.row_dimensions[1].height = 24

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto filter
    ws.auto_filter.ref = 'A1:H1'

    # Page setup: A4, portrait, NO scaling (default 100%)
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    # Do NOT set scale — default is 100% (no scaling applied)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Status Report')
    print(f'  Data rows: 60 (rows 1-60, header + 59 data)')
    print(f'  Columns: A-H (8 columns)')
    print(f'  Print scale: default (100%, no scaling applied)')
    print(f'  Paper: A4, Portrait')


create_initial()
