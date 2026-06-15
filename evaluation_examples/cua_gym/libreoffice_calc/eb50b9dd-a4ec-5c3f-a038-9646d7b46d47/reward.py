"""
Reward Script: Reduce print scale to 75% on 'Status Report' sheet
Task ID: calc_adv_print_scale_pct_023
Domain: libreoffice_calc
Scoring:
  Component 1: Print scale is set to exactly 75 (0.7 pts)
  Component 2: Scale is percentage-based (not fit-to-page) — fitToWidth and fitToHeight are None (0.3 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_print_scale_pct_023'
SHEET_NAME = 'Status Report'
EXPECTED_SCALE = 75


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Set the print scale on sheet 'Status Report' to exactly 75%.
    The scale must be set as a direct percentage (not fit-to-N-pages).
    """
    total_score = 0.0

    # Precondition gate: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: the target sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]
    ps = ws.page_setup

    # Component 1: Print scale is set to exactly 75 (0.7 points)
    # This FAILS on initial (scale=None) and PASSES on golden (scale=75)
    try:
        actual_scale = ps.scale
        if actual_scale == EXPECTED_SCALE:
            print(f"PASS: Component 1 — Print scale is set to {actual_scale}% (expected {EXPECTED_SCALE}%) (0.7 pts)")
            total_score += 0.7
        else:
            print(f"FAIL: Component 1 — Expected print scale={EXPECTED_SCALE}, found scale={actual_scale}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not read page_setup.scale: {e}")

    # Component 2: Scale is percentage-based, NOT fit-to-page (0.3 points)
    # fitToWidth and fitToHeight should both be None (unset), confirming that
    # the scaling was applied as a direct percentage rather than 'fit to N pages'.
    # This FAILS on initial (scale is None, i.e., no scaling was applied at all)
    # and PASSES on golden (scale=75, fitToWidth=None, fitToHeight=None).
    try:
        fit_to_width = ps.fitToWidth
        fit_to_height = ps.fitToHeight
        scale_is_percentage = (ps.scale == EXPECTED_SCALE and fit_to_width is None and fit_to_height is None)
        if scale_is_percentage:
            print(f"PASS: Component 2 — Percentage-based scaling: scale={ps.scale}, fitToWidth={fit_to_width}, fitToHeight={fit_to_height} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Expected percentage scaling (scale=75, fitToWidth=None, fitToHeight=None), found: scale={ps.scale}, fitToWidth={fit_to_width}, fitToHeight={fit_to_height}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not read fit-to-page properties: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
