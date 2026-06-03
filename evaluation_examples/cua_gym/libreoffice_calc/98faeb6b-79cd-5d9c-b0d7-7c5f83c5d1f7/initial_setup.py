"""
Initial Setup: Manufacturing Quality Control Defect Log
Task ID: calc_grs_070
Domain: libreoffice_calc

Creates a defect log spreadsheet with realistic manufacturing data,
data validation dropdowns, and empty placeholder sheets for analysis.
The Defect Rate % column is left empty (task requires building it).
No conditional formatting applied (task requires adding it).
Pareto, Control Chart, and Shift Comparison sheets are empty stubs.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========================================
    # Sheet 1: Defect Log (main data sheet)
    # ========================================
    ws = wb.active
    ws.title = "Defect Log"

    headers = [
        "Defect ID", "Date", "Shift", "Production Line", "Product Code",
        "Defect Type", "Severity", "Quantity Defective", "Batch Size",
        "Defect Rate %", "Inspector", "Disposition"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_widths = {
        "A": 12, "B": 12, "C": 12, "D": 16, "E": 14,
        "F": 14, "G": 12, "H": 18, "I": 12,
        "J": 14, "K": 16, "L": 12,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Realistic defect log data (25 rows)
    defect_data = [
        ["DEF-2025-001", "2025-03-01", "Morning",   "Line A", "PCB-4100", "Surface",      "Minor",    3,  500, None, "Sarah Chen",      "Accept"],
        ["DEF-2025-002", "2025-03-01", "Afternoon",  "Line B", "PCB-4100", "Dimensional",  "Major",    8,  500, None, "Marcus Johnson",   "Rework"],
        ["DEF-2025-003", "2025-03-02", "Night",      "Line C", "MOD-2200", "Assembly",     "Critical", 12, 400, None, "Priya Patel",      "Scrap"],
        ["DEF-2025-004", "2025-03-02", "Morning",    "Line A", "PCB-4100", "Functional",   "Major",    5,  500, None, "Sarah Chen",       "Rework"],
        ["DEF-2025-005", "2025-03-03", "Afternoon",  "Line D", "SEN-3300", "Cosmetic",     "Minor",    2,  600, None, "James Wilson",     "Accept"],
        ["DEF-2025-006", "2025-03-03", "Night",      "Line B", "MOD-2200", "Surface",      "Minor",    4,  400, None, "Yuki Tanaka",      "Accept"],
        ["DEF-2025-007", "2025-03-04", "Morning",    "Line C", "PCB-4100", "Dimensional",  "Major",    15, 500, None, "Marcus Johnson",   "Hold"],
        ["DEF-2025-008", "2025-03-04", "Afternoon",  "Line A", "SEN-3300", "Assembly",     "Critical", 9,  600, None, "Sarah Chen",       "Scrap"],
        ["DEF-2025-009", "2025-03-05", "Night",      "Line D", "MOD-2200", "Functional",   "Major",    7,  400, None, "Priya Patel",      "Rework"],
        ["DEF-2025-010", "2025-03-05", "Morning",    "Line B", "PCB-4100", "Cosmetic",     "Minor",    1,  500, None, "James Wilson",     "Accept"],
        ["DEF-2025-011", "2025-03-06", "Afternoon",  "Line C", "SEN-3300", "Surface",      "Minor",    6,  600, None, "Yuki Tanaka",      "Rework"],
        ["DEF-2025-012", "2025-03-06", "Night",      "Line A", "MOD-2200", "Dimensional",  "Major",    11, 400, None, "Sarah Chen",       "Hold"],
        ["DEF-2025-013", "2025-03-07", "Morning",    "Line D", "PCB-4100", "Assembly",     "Critical", 18, 500, None, "Marcus Johnson",   "Scrap"],
        ["DEF-2025-014", "2025-03-07", "Afternoon",  "Line B", "SEN-3300", "Functional",   "Minor",    3,  600, None, "Priya Patel",      "Accept"],
        ["DEF-2025-015", "2025-03-08", "Night",      "Line C", "MOD-2200", "Cosmetic",     "Minor",    2,  400, None, "James Wilson",     "Accept"],
        ["DEF-2025-016", "2025-03-08", "Morning",    "Line A", "PCB-4100", "Surface",      "Major",    10, 500, None, "Yuki Tanaka",      "Rework"],
        ["DEF-2025-017", "2025-03-09", "Afternoon",  "Line D", "SEN-3300", "Dimensional",  "Major",    14, 600, None, "Sarah Chen",       "Hold"],
        ["DEF-2025-018", "2025-03-09", "Night",      "Line B", "MOD-2200", "Assembly",     "Critical", 20, 400, None, "Marcus Johnson",   "Scrap"],
        ["DEF-2025-019", "2025-03-10", "Morning",    "Line C", "PCB-4100", "Functional",   "Minor",    4,  500, None, "Priya Patel",      "Accept"],
        ["DEF-2025-020", "2025-03-10", "Afternoon",  "Line A", "SEN-3300", "Cosmetic",     "Minor",    1,  600, None, "James Wilson",     "Accept"],
        ["DEF-2025-021", "2025-03-11", "Night",      "Line D", "MOD-2200", "Surface",      "Major",    8,  400, None, "Yuki Tanaka",      "Rework"],
        ["DEF-2025-022", "2025-03-11", "Morning",    "Line B", "PCB-4100", "Dimensional",  "Critical", 16, 500, None, "Sarah Chen",       "Scrap"],
        ["DEF-2025-023", "2025-03-12", "Afternoon",  "Line C", "SEN-3300", "Assembly",     "Major",    7,  600, None, "Marcus Johnson",   "Rework"],
        ["DEF-2025-024", "2025-03-12", "Night",      "Line A", "MOD-2200", "Functional",   "Minor",    3,  400, None, "Priya Patel",      "Accept"],
        ["DEF-2025-025", "2025-03-13", "Morning",    "Line D", "PCB-4100", "Cosmetic",     "Minor",    2,  500, None, "James Wilson",     "Accept"],
    ]

    data_align = Alignment(horizontal="center", vertical="center")
    for r, row_data in enumerate(defect_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = data_align
            # Date format
            if c == 2 and val:
                cell.number_format = 'yyyy-mm-dd'

    # Defect Rate % column (J) intentionally left empty/None - task requires formula
    # No conditional formatting applied - task requires adding it

    # --- Data Validations (Dropdowns) ---
    # Shift
    dv_shift = DataValidation(
        type="list",
        formula1='"Morning,Afternoon,Night"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_shift.prompt = "Select shift"
    dv_shift.promptTitle = "Shift"
    dv_shift.add("C2:C100")
    ws.add_data_validation(dv_shift)

    # Production Line
    dv_line = DataValidation(
        type="list",
        formula1='"Line A,Line B,Line C,Line D"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_line.prompt = "Select production line"
    dv_line.promptTitle = "Production Line"
    dv_line.add("D2:D100")
    ws.add_data_validation(dv_line)

    # Defect Type
    dv_defect = DataValidation(
        type="list",
        formula1='"Dimensional,Surface,Assembly,Functional,Cosmetic"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_defect.prompt = "Select defect type"
    dv_defect.promptTitle = "Defect Type"
    dv_defect.add("F2:F100")
    ws.add_data_validation(dv_defect)

    # Severity
    dv_severity = DataValidation(
        type="list",
        formula1='"Critical,Major,Minor"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_severity.prompt = "Select severity"
    dv_severity.promptTitle = "Severity"
    dv_severity.add("G2:G100")
    ws.add_data_validation(dv_severity)

    # Disposition
    dv_disposition = DataValidation(
        type="list",
        formula1='"Rework,Scrap,Accept,Hold"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_disposition.prompt = "Select disposition"
    dv_disposition.promptTitle = "Disposition"
    dv_disposition.add("L2:L100")
    ws.add_data_validation(dv_disposition)

    # Freeze header row
    ws.freeze_panes = "A2"

    # ========================================
    # Sheet 2: Pareto Analysis (empty stub)
    # ========================================
    ws_pareto = wb.create_sheet("Pareto Analysis")
    ws_pareto["A1"] = "Defect Type"
    ws_pareto["B1"] = "Count"
    ws_pareto["C1"] = "Cumulative %"
    for col in range(1, 4):
        cell = ws_pareto.cell(row=1, column=col)
        cell.font = Font(bold=True)

    # ========================================
    # Sheet 3: Control Chart (empty stub)
    # ========================================
    ws_control = wb.create_sheet("Control Chart")
    ws_control["A1"] = "Date"
    ws_control["B1"] = "Defect Rate %"
    for col in range(1, 3):
        cell = ws_control.cell(row=1, column=col)
        cell.font = Font(bold=True)

    # ========================================
    # Sheet 4: Shift Comparison (empty stub)
    # ========================================
    ws_shift = wb.create_sheet("Shift Comparison")
    ws_shift["A1"] = "Shift"
    ws_shift["B1"] = "Total Defects"
    ws_shift["C1"] = "Avg Defect Rate %"
    for col in range(1, 4):
        cell = ws_shift.cell(row=1, column=col)
        cell.font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
