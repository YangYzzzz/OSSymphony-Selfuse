"""
Reward Script: Manufacturing Quality Control Defect Log
Task ID: calc_grs_070
Domain: libreoffice_calc
Scoring:
  Component 1: Defect Rate formulas in J column (0.15)
  Component 2: Defect Rate % number format (0.05)
  Component 3: Conditional formatting on Defect Rate (0.15)
  Component 4: Pareto Analysis data populated & sorted descending (0.20)
  Component 5: Pareto Analysis chart exists (0.10)
  Component 6: Control Chart data populated (0.10)
  Component 7: Control Chart (line chart) exists (0.10)
  Component 8: Shift Comparison data populated (0.10)
  Component 9: Shift Comparison chart exists (0.05)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_070'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets must exist
    required_sheets = ['Defect Log', 'Pareto Analysis', 'Control Chart', 'Shift Comparison']
    for sn in required_sheets:
        if sn not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sn}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    ws_log = wb['Defect Log']
    ws_pareto = wb['Pareto Analysis']
    ws_control = wb['Control Chart']
    ws_shift = wb['Shift Comparison']

    # Component 1: Defect Rate formulas in J column (0.15 points)
    # Golden has =H{row}/I{row} formulas in J2:J26; initial has None
    try:
        formula_count = 0
        data_rows = 0
        for row in range(2, ws_log.max_row + 1):
            h_val = ws_log.cell(row=row, column=8).value
            if h_val is not None:
                data_rows += 1
                j_val = ws_log.cell(row=row, column=10).value
                if j_val is not None and isinstance(j_val, str) and '/' in j_val.upper():
                    formula_count += 1
                elif j_val is not None and isinstance(j_val, str) and '=' in j_val:
                    # Any formula referencing H and I columns counts
                    if 'H' in j_val.upper() and 'I' in j_val.upper():
                        formula_count += 1

        if data_rows > 0 and formula_count >= data_rows * 0.8:
            print(f"PASS: Component 1 — Defect Rate formulas found in {formula_count}/{data_rows} data rows (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Expected formulas in J column, found {formula_count}/{data_rows} rows with formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Defect Rate % number format (0.05 points)
    # Golden has 0.00% format; initial has General
    try:
        j2_fmt = ws_log['J2'].number_format
        if j2_fmt is not None and '%' in str(j2_fmt):
            print(f"PASS: Component 2 — J2 number format is '{j2_fmt}' (contains %) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — Expected percentage format in J2, found: '{j2_fmt}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on Defect Rate column (0.15 points)
    # Golden has 3 CF rules on J column (>5% red, 2-5% yellow, <2% green); initial has 0
    try:
        cf_rules_on_j = 0
        for cf in ws_log.conditional_formatting:
            cf_range = str(cf)
            if 'J' in cf_range:
                cf_rules_on_j += len(cf.rules)

        if cf_rules_on_j >= 3:
            print(f"PASS: Component 3 — {cf_rules_on_j} conditional formatting rules on J column (0.15 pts)")
            total_score += 0.15
        elif cf_rules_on_j >= 1:
            partial = 0.05 * cf_rules_on_j
            print(f"PARTIAL: Component 3 — {cf_rules_on_j}/3 conditional formatting rules on J column ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found on J column")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Pareto Analysis data populated and sorted descending (0.20 points)
    # Golden has 5+ data rows with defect types, counts sorted descending, and cumulative %
    # Initial has only headers (1 row)
    try:
        pareto_data_rows = ws_pareto.max_row - 1 if ws_pareto.max_row else 0
        # Check if there are actual data rows (not just headers)
        actual_data = 0
        for row in range(2, ws_pareto.max_row + 1):
            if ws_pareto.cell(row=row, column=1).value is not None:
                actual_data += 1

        if actual_data >= 3:
            # Check if sorted descending by count (column B)
            counts = []
            for row in range(2, ws_pareto.max_row + 1):
                b_val = ws_pareto.cell(row=row, column=2).value
                if b_val is not None:
                    try:
                        counts.append(float(b_val))
                    except (ValueError, TypeError):
                        pass

            is_sorted_desc = all(counts[i] >= counts[i+1] for i in range(len(counts)-1)) if len(counts) > 1 else False

            # Check for cumulative percentage column
            cumulative_ok = any(
                ws_pareto.cell(row=1, column=col).value is not None
                and 'cumul' in str(ws_pareto.cell(row=1, column=col).value).lower()
                and ws_pareto.cell(row=actual_data + 1, column=col).value is not None
                and 95 <= float(ws_pareto.cell(row=actual_data + 1, column=col).value) <= 100.1
                for col in range(1, ws_pareto.max_column + 1)
                if ws_pareto.cell(row=1, column=col).value is not None
                and 'cumul' in str(ws_pareto.cell(row=1, column=col).value).lower()
            )

            if is_sorted_desc and cumulative_ok:
                print(f"PASS: Component 4 — Pareto has {actual_data} data rows, sorted descending, with cumulative % (0.20 pts)")
                total_score += 0.20
            elif is_sorted_desc:
                print(f"PARTIAL: Component 4 — Pareto sorted descending but missing/incorrect cumulative % (0.12 pts)")
                total_score += 0.12
            elif actual_data >= 3:
                print(f"PARTIAL: Component 4 — Pareto has {actual_data} rows but not sorted descending (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 4 — Insufficient Pareto data ({actual_data} rows)")
        else:
            print(f"FAIL: Component 4 — Pareto Analysis has only {actual_data} data rows (need >= 3)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Pareto Analysis chart exists (0.10 points)
    # Golden has 1 BarChart; initial has 0
    try:
        pareto_charts = ws_pareto._charts
        if len(pareto_charts) >= 1:
            print(f"PASS: Component 5 — Pareto Analysis has {len(pareto_charts)} chart(s) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No charts in Pareto Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Control Chart data populated (0.10 points)
    # Golden has 13+ data rows with dates and defect rates; initial has only headers
    try:
        control_data = 0
        for row in range(2, ws_control.max_row + 1):
            if ws_control.cell(row=row, column=1).value is not None:
                control_data += 1

        if control_data >= 5:
            print(f"PASS: Component 6 — Control Chart has {control_data} data rows (0.10 pts)")
            total_score += 0.10
        elif control_data >= 2:
            print(f"PARTIAL: Component 6 — Control Chart has only {control_data} data rows (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Control Chart has only {control_data} data rows (need >= 5)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Control Chart (line chart) exists (0.10 points)
    # Golden has 1 LineChart; initial has 0
    try:
        control_charts = ws_control._charts
        if len(control_charts) >= 1:
            print(f"PASS: Component 7 — Control Chart has {len(control_charts)} chart(s) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 — No charts in Control Chart sheet")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: Shift Comparison data populated (0.10 points)
    # Golden has 3 rows (Morning, Afternoon, Night) with aggregated stats; initial has only headers
    try:
        shift_data = 0
        shift_names = set()
        for row in range(2, ws_shift.max_row + 1):
            val = ws_shift.cell(row=row, column=1).value
            if val is not None:
                shift_data += 1
                shift_names.add(str(val).strip())

        expected_shifts = {'Morning', 'Afternoon', 'Night'}
        if shift_data >= 3 and expected_shifts.issubset(shift_names):
            print(f"PASS: Component 8 — Shift Comparison has all 3 shifts with data (0.10 pts)")
            total_score += 0.10
        elif shift_data >= 2:
            print(f"PARTIAL: Component 8 — Shift Comparison has {shift_data} rows, shifts: {shift_names} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 8 — Shift Comparison has only {shift_data} data rows")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    # Component 9: Shift Comparison chart exists (0.05 points)
    # Golden has 1 BarChart; initial has 0
    try:
        shift_charts = ws_shift._charts
        if len(shift_charts) >= 1:
            print(f"PASS: Component 9 — Shift Comparison has {len(shift_charts)} chart(s) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 9 — No charts in Shift Comparison sheet")
    except Exception as e:
        print(f"ERROR: Component 9 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
