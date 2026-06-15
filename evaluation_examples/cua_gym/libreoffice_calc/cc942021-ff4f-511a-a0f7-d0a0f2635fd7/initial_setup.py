"""
Initial Setup: Create annual budget spreadsheet with empty headers/footers
Task ID: calc_gao_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gao_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Budget ---
    ws = wb.active
    ws.title = 'Budget'

    # Column headers (A through K)
    headers = [
        'Category', 'Department', 'Q1 Budget', 'Q1 Actual', 'Q2 Budget',
        'Q2 Actual', 'Q3 Budget', 'Q3 Actual', 'Q4 Budget', 'Q4 Actual', 'Annual Total'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths
    col_widths = {'A': 22, 'B': 18, 'C': 14, 'D': 14, 'E': 14, 'F': 14,
                  'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 16}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row height for header
    ws.row_dimensions[1].height = 24

    # Budget data - realistic annual budget for a mid-size company
    departments = ['Engineering', 'Marketing', 'Sales', 'Human Resources',
                   'Finance', 'Operations', 'Legal', 'Customer Support',
                   'Research', 'Executive']

    categories_by_dept = {
        'Engineering': ['Salaries', 'Software Licenses', 'Hardware', 'Cloud Infrastructure',
                        'Training & Development', 'Contractor Fees', 'Office Supplies', 'Travel'],
        'Marketing': ['Salaries', 'Digital Advertising', 'Events & Conferences', 'Content Production',
                      'PR Agency', 'Market Research', 'Brand Development', 'Social Media Tools'],
        'Sales': ['Salaries', 'Commission', 'CRM Software', 'Lead Generation',
                  'Client Entertainment', 'Travel', 'Sales Materials', 'Training'],
        'Human Resources': ['Salaries', 'Recruiting', 'Benefits Administration', 'Employee Engagement',
                            'Training Programs', 'HR Software', 'Legal Compliance', 'Wellness Programs'],
        'Finance': ['Salaries', 'Audit Fees', 'Accounting Software', 'Insurance',
                    'Tax Consulting', 'Banking Fees', 'Financial Reporting', 'Compliance'],
        'Operations': ['Salaries', 'Facility Maintenance', 'Utilities', 'Security',
                       'Equipment Leases', 'Shipping & Logistics', 'Warehouse', 'Fleet Management'],
        'Legal': ['Salaries', 'Outside Counsel', 'Patent Filings', 'Contract Review',
                  'Regulatory Compliance', 'Litigation Reserve', 'Legal Software', 'IP Protection'],
        'Customer Support': ['Salaries', 'Help Desk Software', 'Training', 'Quality Assurance',
                             'Phone Systems', 'Customer Feedback Tools', 'Knowledge Base', 'Outsourcing'],
        'Research': ['Salaries', 'Lab Equipment', 'Materials & Supplies', 'Patent Applications',
                     'Conferences', 'Publications', 'Collaboration Tools', 'External Research'],
        'Executive': ['Salaries', 'Board Expenses', 'Strategic Consulting', 'Investor Relations',
                      'Corporate Development', 'Executive Travel', 'Memberships', 'Philanthropy'],
    }

    # Budget values (in thousands) - somewhat realistic
    import random
    random.seed(2024046)

    money_fmt = '#,##0'
    data_font = Font(name='Calibri', size=10)
    data_align_text = Alignment(horizontal='left', vertical='center')
    data_align_num = Alignment(horizontal='right', vertical='center')

    row = 2
    for dept in departments:
        cats = categories_by_dept[dept]
        for cat in cats:
            # Generate budget values
            base = random.randint(15, 350) * 1000
            q1_budget = base + random.randint(-5000, 5000)
            q1_actual = int(q1_budget * random.uniform(0.85, 1.15))
            q2_budget = base + random.randint(-8000, 8000)
            q2_actual = int(q2_budget * random.uniform(0.88, 1.12))
            q3_budget = base + random.randint(-6000, 6000)
            q3_actual = int(q3_budget * random.uniform(0.82, 1.18))
            q4_budget = base + random.randint(-7000, 7000)
            q4_actual = int(q4_budget * random.uniform(0.90, 1.10))
            annual_total = q1_actual + q2_actual + q3_actual + q4_actual

            row_data = [cat, dept, q1_budget, q1_actual, q2_budget,
                        q2_actual, q3_budget, q3_actual, q4_budget, q4_actual, annual_total]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                cell.border = thin_border
                cell.font = data_font
                if col_idx <= 2:
                    cell.alignment = data_align_text
                else:
                    cell.alignment = data_align_num
                    cell.number_format = money_fmt

            row += 1

    # Fill remaining rows to reach at least 80 rows of content
    # We have 10 departments * 8 categories = 80 rows + 1 header = row 81
    # That gives us data from row 2 to row 81, spanning A1:K81

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Headers and footers are intentionally left EMPTY (task requires setting them)
    # openpyxl creates worksheets with no headers/footers by default

    # Set print area to ensure multi-page print
    ws.print_area = 'A1:K81'

    # Page setup for printing
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.fitToPage = True

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
