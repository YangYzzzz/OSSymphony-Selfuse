"""
Initial Setup: Product Pricing Analysis Spreadsheet
Task ID: calc_grs_034
Domain: libreoffice_calc

Creates a spreadsheet with 20 products containing raw data columns:
Product Name, Cost Price, Desired Margin %, Competitor Price.
Columns for Selling Price, Price Difference, and Price Position are present but EMPTY.
A What-If analysis section header is placed below, but no formulas or conditional
formatting or charts are included.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Pricing"

    # --- Headers (Row 1) ---
    headers = [
        "Product Name",
        "Cost Price",
        "Desired Margin %",
        "Calculated Selling Price",
        "Competitor Price",
        "Price Difference",
        "Price Position",
    ]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # --- Product Data (Rows 2-21) ---
    # Columns: Product Name, Cost Price, Desired Margin %, (empty), Competitor Price, (empty), (empty)
    products = [
        ("Wireless Bluetooth Earbuds",        24.50,  0.35,  35.99),
        ("USB-C Charging Cable 6ft",           3.80,  0.45,   6.49),
        ("Stainless Steel Water Bottle",       8.20,  0.40,  12.99),
        ("Bamboo Phone Stand",                 5.60,  0.50,   9.99),
        ("LED Desk Lamp",                     18.30,  0.38,  27.50),
        ("Portable Power Bank 10000mAh",      15.75,  0.42,  24.99),
        ("Microfiber Cleaning Cloth Set",      2.40,  0.55,   4.49),
        ("Ergonomic Mouse Pad",                6.90,  0.35,  10.99),
        ("Laptop Privacy Screen 15in",        22.00,  0.40,  34.99),
        ("Noise Cancelling Headphones",       45.00,  0.30,  62.00),
        ("Mechanical Keyboard",               38.50,  0.35,  54.99),
        ("Webcam HD 1080p",                   19.80,  0.40,  29.99),
        ("Monitor Riser Stand",               14.20,  0.38,  21.50),
        ("Cable Management Kit",               7.50,  0.48,  12.99),
        ("Surge Protector Power Strip",       11.40,  0.36,  16.99),
        ("Desk Organizer Tray",                9.80,  0.42,  15.49),
        ("Wireless Charging Pad",             12.60,  0.35,  18.99),
        ("USB Hub 7-Port",                    16.90,  0.40,  25.99),
        ("Anti-Fatigue Standing Mat",         28.00,  0.32,  39.99),
        ("Screen Cleaning Spray Kit",          4.50,  0.50,   7.99),
    ]

    for r, (name, cost, margin, competitor) in enumerate(products, 2):
        ws.cell(row=r, column=1, value=name).border = thin_border
        ws.cell(row=r, column=2, value=cost).border = thin_border
        ws.cell(row=r, column=2).number_format = '$#,##0.00'
        ws.cell(row=r, column=3, value=margin).border = thin_border
        ws.cell(row=r, column=3).number_format = '0%'
        # Column 4 (Calculated Selling Price) - EMPTY (task asks agent to add formula)
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=4).number_format = '$#,##0.00'
        ws.cell(row=r, column=5, value=competitor).border = thin_border
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        # Column 6 (Price Difference) - EMPTY
        ws.cell(row=r, column=6).border = thin_border
        ws.cell(row=r, column=6).number_format = '$#,##0.00'
        # Column 7 (Price Position) - EMPTY
        ws.cell(row=r, column=7).border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18

    # --- What-If Analysis Section (below data, row 24+) ---
    ws.cell(row=24, column=1, value="What-If Analysis").font = Font(
        name="Calibri", size=14, bold=True
    )
    ws.cell(row=25, column=1, value="Desired Margin %:")
    ws.cell(row=25, column=1).font = Font(name="Calibri", size=11, bold=True)
    ws.cell(row=25, column=2, value=0.35)
    ws.cell(row=25, column=2).number_format = '0%'
    ws.cell(row=25, column=2).border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="medium", color="000000"),
        bottom=Side(style="medium", color="000000"),
    )
    ws.cell(row=25, column=2).fill = PatternFill(
        start_color="FFFFF2CC", end_color="FFFFF2CC", fill_type="solid"
    )

    # What-If table headers (row 27)
    whatif_headers = ["Product Name", "Cost Price", "What-If Selling Price", "What-If Competitor Diff"]
    for col_idx, h in enumerate(whatif_headers, 1):
        cell = ws.cell(row=27, column=col_idx, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
        cell.border = thin_border

    # What-If data rows (28-47): product names and cost prices copied, formula columns EMPTY
    for r, (name, cost, margin, competitor) in enumerate(products, 28):
        ws.cell(row=r, column=1, value=name).border = thin_border
        ws.cell(row=r, column=2, value=cost).border = thin_border
        ws.cell(row=r, column=2).number_format = '$#,##0.00'
        # Columns 3 and 4 are EMPTY (agent adds What-If formulas)
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=3).number_format = '$#,##0.00'
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=4).number_format = '$#,##0.00'

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
