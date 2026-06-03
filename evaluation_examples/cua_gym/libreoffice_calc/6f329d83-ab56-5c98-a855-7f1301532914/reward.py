"""
Reward Script: Product Pricing Analysis Spreadsheet
Task ID: calc_grs_034
Domain: libreoffice_calc
Scoring:
  Component 1: Selling Price formulas in D2:D21 (0.20 pts)
  Component 2: Price Difference formulas in F2:F21 (0.15 pts)
  Component 3: Price Position IF formulas in G2:G21 (0.20 pts)
  Component 4: What-If Selling Price formulas in C28:C47 referencing $B$25 (0.15 pts)
  Component 5: What-If Competitor Diff formulas in D28:D47 (0.10 pts)
  Component 6: Conditional formatting on G2:G21 for Below/Above Market (0.10 pts)
  Component 7: Scatter chart present with Cost Price vs Selling Price (0.10 pts)
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_034'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet exists
    if 'Product Pricing' not in wb.sheetnames:
        print("CRITICAL: 'Product Pricing' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Product Pricing']

    # Component 1: Selling Price formulas in D2:D21 (0.20 points)
    # Formula pattern: =B<row>*(1+C<row>) — cost * (1 + margin)
    try:
        formula_count = 0
        for row in range(2, 22):
            val = ws.cell(row=row, column=4).value  # Column D
            if val and isinstance(val, str) and val.startswith('='):
                # Check it references B and C in same row for margin calc
                val_upper = val.upper().replace(' ', '')
                if ('B' in val_upper and 'C' in val_upper) or ('1+' in val_upper or '(1+' in val_upper):
                    formula_count += 1
        if formula_count >= 18:  # Allow up to 2 slightly different
            print(f"PASS: Component 1 — Selling Price formulas found in {formula_count}/20 rows (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 1 — Selling Price formulas in {formula_count}/20 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected selling price formulas in D2:D21, found {formula_count} formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Price Difference formulas in F2:F21 (0.15 points)
    # Formula pattern: =D<row>-E<row> — selling price minus competitor price
    try:
        formula_count = 0
        for row in range(2, 22):
            val = ws.cell(row=row, column=6).value  # Column F
            if val and isinstance(val, str) and val.startswith('='):
                val_upper = val.upper().replace(' ', '')
                # Should reference D and E (selling price - competitor price)
                if 'D' in val_upper and 'E' in val_upper:
                    formula_count += 1
        if formula_count >= 18:
            print(f"PASS: Component 2 — Price Difference formulas found in {formula_count}/20 rows (0.15 pts)")
            total_score += 0.15
        elif formula_count >= 10:
            partial = 0.08
            print(f"PARTIAL: Component 2 — Price Difference formulas in {formula_count}/20 rows ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected price diff formulas in F2:F21, found {formula_count} formulas")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Price Position IF formulas in G2:G21 (0.20 points)
    # Should use IF to classify as Above/At/Below Market based on tolerance
    try:
        formula_count = 0
        has_market_text = 0
        for row in range(2, 22):
            val = ws.cell(row=row, column=7).value  # Column G
            if val and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                if val_upper.startswith('=') and 'IF' in val_upper:
                    formula_count += 1
                    if 'MARKET' in val_upper or 'market' in val.lower():
                        has_market_text += 1
        if formula_count >= 18 and has_market_text >= 18:
            print(f"PASS: Component 3 — Price Position IF formulas with Market labels in {formula_count}/20 rows (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 3 — IF formulas in {formula_count}/20 rows, {has_market_text} with Market text ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected IF formulas in G2:G21, found {formula_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: What-If Selling Price formulas in C28:C47 referencing $B$25 (0.15 points)
    # Formula pattern: =B<row>*(1+$B$25) — uses the what-if margin input cell
    try:
        formula_count = 0
        refs_b25 = 0
        for row in range(28, 48):
            val = ws.cell(row=row, column=3).value  # Column C
            if val and isinstance(val, str) and val.startswith('='):
                formula_count += 1
                val_upper = val.upper().replace(' ', '')
                if '$B$25' in val_upper or 'B$25' in val_upper or '$B25' in val_upper:
                    refs_b25 += 1
        if formula_count >= 18 and refs_b25 >= 18:
            print(f"PASS: Component 4 — What-If formulas in C28:C47 with $B$25 ref: {refs_b25}/20 (0.15 pts)")
            total_score += 0.15
        elif formula_count >= 10:
            partial = 0.08
            print(f"PARTIAL: Component 4 — {formula_count} formulas, {refs_b25} ref $B$25 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Expected What-If formulas in C28:C47, found {formula_count}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: What-If Competitor Diff formulas in D28:D47 (0.10 points)
    try:
        formula_count = 0
        for row in range(28, 48):
            val = ws.cell(row=row, column=4).value  # Column D in what-if section
            if val and isinstance(val, str) and val.startswith('='):
                formula_count += 1
        if formula_count >= 18:
            print(f"PASS: Component 5 — What-If Competitor Diff formulas in {formula_count}/20 rows (0.10 pts)")
            total_score += 0.10
        elif formula_count >= 10:
            partial = 0.05
            print(f"PARTIAL: Component 5 — {formula_count}/20 What-If diff formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Expected What-If diff formulas in D28:D47, found {formula_count}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on G column for Below/Above Market (0.10 points)
    try:
        cf_rules = list(ws.conditional_formatting)
        has_below_market_rule = False
        has_above_market_rule = False
        for cf in cf_rules:
            cf_range = str(cf)
            # Check if the range covers column G in the data area
            if 'G' in cf_range:
                for rule in cf.rules:
                    formula_list = getattr(rule, 'formula', [])
                    for f in formula_list:
                        f_upper = f.upper().replace(' ', '')
                        if 'BELOWMARKET' in f_upper or 'BELOW MARKET' in f.upper():
                            has_below_market_rule = True
                        if 'ABOVEMARKET' in f_upper or 'ABOVE MARKET' in f.upper():
                            has_above_market_rule = True
                    # Also check operator-based rules
                    op = getattr(rule, 'operator', None)
                    if op == 'equal' and formula_list:
                        for f in formula_list:
                            if 'Below Market' in f or 'below market' in f.lower():
                                has_below_market_rule = True
                            if 'Above Market' in f or 'above market' in f.lower():
                                has_above_market_rule = True

        if has_below_market_rule and has_above_market_rule:
            print(f"PASS: Component 6 — Conditional formatting for Below Market and Above Market (0.10 pts)")
            total_score += 0.10
        elif has_below_market_rule or has_above_market_rule:
            partial = 0.05
            found = 'Below Market' if has_below_market_rule else 'Above Market'
            print(f"PARTIAL: Component 6 — Only {found} rule found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — No conditional formatting rules for Below/Above Market on column G")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Scatter chart present (0.10 points)
    # Task: scatter plot with Cost Price on X-axis and Selling Price on Y-axis
    try:
        from openpyxl.chart import ScatterChart as ScatterChartClass
        charts = ws._charts
        scatter_found = False
        for chart in charts:
            if isinstance(chart, ScatterChartClass):
                scatter_found = True
                # Check if it has series data
                if len(chart.series) >= 1:
                    print(f"PASS: Component 7 — Scatter chart found with {len(chart.series)} series (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"PARTIAL: Component 7 — Scatter chart found but no series data (0.05 pts)")
                    total_score += 0.05
                break
        if not scatter_found:
            # Check if any chart exists (might not be ScatterChart class)
            if len(charts) > 0:
                print(f"PARTIAL: Component 7 — Chart found but not a scatter chart (0.03 pts)")
                total_score += 0.03
            else:
                print(f"FAIL: Component 7 — No scatter chart found")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
