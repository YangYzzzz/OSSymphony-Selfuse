"""
Reward Script: Convert text dates in column B to proper date values
Task ID: calc_tbl_071
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5) - Column B cells B2:B16 contain datetime objects, not text
  Component 2 (0.3) - Date values are correct (match expected dates)
  Component 3 (0.2) - SUMIFS formula in B22 is preserved
"""

import os
import datetime

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_071'

# Expected date values for B2:B16 (the ground truth dates)
EXPECTED_DATES = {
    2: datetime.datetime(2023, 11, 10),
    3: datetime.datetime(2023, 12, 5),
    4: datetime.datetime(2024, 1, 15),
    5: datetime.datetime(2024, 1, 22),
    6: datetime.datetime(2024, 2, 8),
    7: datetime.datetime(2024, 2, 14),
    8: datetime.datetime(2024, 3, 1),
    9: datetime.datetime(2024, 3, 18),
    10: datetime.datetime(2024, 4, 12),
    11: datetime.datetime(2024, 4, 25),
    12: datetime.datetime(2024, 5, 3),
    13: datetime.datetime(2024, 5, 19),
    14: datetime.datetime(2024, 6, 7),
    15: datetime.datetime(2024, 6, 22),
    16: datetime.datetime(2024, 7, 15),
}


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active
    if ws is None:
        ws = wb.worksheets[0]

    # Component 1: Column B cells B2:B16 are datetime, not text strings (0.5 points)
    # This is the core task: converting text dates to proper date values.
    # In the initial state, all B2:B16 are str; in golden state, all are datetime.
    try:
        datetime_count = 0
        total_date_cells = len(EXPECTED_DATES)  # 15 cells
        for row_num in EXPECTED_DATES:
            cell = ws.cell(row=row_num, column=2)
            val = cell.value
            if isinstance(val, datetime.datetime):
                datetime_count += 1
            elif isinstance(val, datetime.date) and not isinstance(val, datetime.datetime):
                # date objects (not datetime) also count as proper dates
                datetime_count += 1

        if datetime_count == total_date_cells:
            print(f"PASS: Component 1 — All {total_date_cells} date cells are datetime type (0.5 pts)")
            total_score += 0.5
        elif datetime_count > 0:
            partial = 0.5 * (datetime_count / total_date_cells)
            if partial > 0:
                print(f"PARTIAL: Component 1 — {datetime_count}/{total_date_cells} cells are datetime ({partial:.3f} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 1 — 0/{total_date_cells} cells are datetime (all still text)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Date values are correct (0.3 points)
    # Each converted date must match the expected date value.
    # Only award points if the cell IS a date (text cells that happen to match don't count).
    try:
        correct_count = 0
        for row_num, expected_dt in EXPECTED_DATES.items():
            cell = ws.cell(row=row_num, column=2)
            val = cell.value
            if isinstance(val, (datetime.datetime, datetime.date)):
                # Compare date portion only (ignore time component)
                if val.year == expected_dt.year and val.month == expected_dt.month and val.day == expected_dt.day:
                    correct_count += 1

        if correct_count == total_date_cells:
            print(f"PASS: Component 2 — All {total_date_cells} date values are correct (0.3 pts)")
            total_score += 0.3
        elif correct_count > 0:
            partial = 0.3 * (correct_count / total_date_cells)
            if partial > 0:
                print(f"PARTIAL: Component 2 — {correct_count}/{total_date_cells} correct date values ({partial:.3f} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 2 — No cells have correct date values (or none are datetime type)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: SUMIFS formula in B22 is preserved (0.2 points)
    # The formula should still exist after the conversion — the agent should not have removed it.
    # This checks that the formula is intact. In both initial and golden, the formula exists,
    # but only scores points when combined with the date conversion (Components 1+2).
    # However, since this check alone would pass on initial too, we gate it:
    # Award points only if at least some dates were converted (Component 1 gave > 0 points).
    try:
        b22_val = ws.cell(row=22, column=2).value
        formula_present = isinstance(b22_val, str) and 'SUMIFS' in b22_val.upper()
        if formula_present and datetime_count > 0:
            print(f"PASS: Component 3 — SUMIFS formula preserved in B22: {b22_val} (0.2 pts)")
            total_score += 0.2
        elif formula_present and datetime_count == 0:
            print(f"FAIL: Component 3 — SUMIFS exists but no dates converted (gated, 0 pts)")
        else:
            print(f"FAIL: Component 3 — SUMIFS formula missing or modified in B22, found: {b22_val!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state("libreoffice_calc")
    verify_task(file_path)
