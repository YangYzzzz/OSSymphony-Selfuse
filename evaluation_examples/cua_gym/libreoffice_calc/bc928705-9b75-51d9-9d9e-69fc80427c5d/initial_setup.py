"""
Initial Setup: Sales data with text dates causing SUMIFS to return 0
Task ID: calc_tbl_071
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Sales Data ---
    ws = wb.active
    ws.title = 'Sales Data'

    # Headers
    headers = ['Department', 'Date', 'Amount']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - dates are TEXT strings (the core problem)
    # Mix of Sales and other departments, dates before and after 2024-01-01
    data = [
        ['Sales',       '2023-11-10',  4520.00],
        ['Marketing',   '2023-12-05',  3200.00],
        ['Sales',       '2024-01-15',  8750.00],
        ['Engineering', '2024-01-22',  6100.00],
        ['Sales',       '2024-02-08',  12340.00],
        ['Marketing',   '2024-02-14',  2890.00],
        ['Sales',       '2024-03-01',  9500.00],
        ['Engineering', '2024-03-18',  7200.00],
        ['Sales',       '2024-04-12',  6830.00],
        ['Marketing',   '2024-04-25',  4100.00],
        ['Sales',       '2024-05-03',  11200.00],
        ['Engineering', '2024-05-19',  5450.00],
        ['Sales',       '2024-06-07',  7650.00],
        ['Marketing',   '2024-06-22',  3750.00],
        ['Sales',       '2024-07-15',  14200.00],
    ]

    for r, row_data in enumerate(data, 2):
        dept_cell = ws.cell(row=r, column=1, value=row_data[0])
        dept_cell.border = thin_border

        # CRITICAL: Dates as TEXT strings - prefix with apostrophe-like approach
        # In openpyxl, setting a string value stores it as text
        date_cell = ws.cell(row=r, column=2, value=row_data[1])
        date_cell.border = thin_border
        date_cell.alignment = Alignment(horizontal='center')
        # Explicitly keep as text format
        date_cell.number_format = '@'

        amount_cell = ws.cell(row=r, column=3, value=row_data[2])
        amount_cell.number_format = '$#,##0.00'
        amount_cell.border = thin_border

    # --- Summary Section ---
    ws.cell(row=18, column=1, value='Summary').font = Font(bold=True, size=12)

    ws.cell(row=19, column=1, value='Department:')
    ws.cell(row=19, column=2, value='Sales')
    ws.cell(row=19, column=2).font = Font(bold=True)

    ws.cell(row=20, column=1, value='Date On or After:')
    ws.cell(row=20, column=2, value='2024-01-01')
    ws.cell(row=20, column=2).number_format = '@'

    ws.cell(row=22, column=1, value='SUMIFS Result:')
    ws.cell(row=22, column=1).font = Font(bold=True)
    # This SUMIFS checks for "Sales" in col A AND dates >= 2024-01-01 in col B
    # It returns 0 because col B dates are text, not real dates
    formula_cell = ws.cell(row=22, column=2,
                           value='=SUMIFS(C2:C16,A2:A16,"Sales",B2:B16,">="&DATE(2024,1,1))')
    formula_cell.number_format = '$#,##0.00'
    formula_cell.font = Font(bold=True, color='FF0000')

    ws.cell(row=23, column=1, value='(Returns 0 because dates in column B are stored as text)')
    ws.cell(row=23, column=1).font = Font(italic=True, color='808080')

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
