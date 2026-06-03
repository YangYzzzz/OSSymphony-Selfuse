"""
Reward Script: INDEX/MATCH/MATCH formula verification
Task ID: calc_lf_003
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): I2 contains a formula (starts with '=')
  Component 2 (0.3): Formula uses INDEX + MATCH pattern for two-way lookup
  Component 3 (0.3): Formula is functionally correct (correct ranges, would yield 4.25)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_003'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Rates' sheet must exist
    if 'Rates' not in wb.sheetnames:
        print("FAIL: 'Rates' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Rates']

    # Get the raw value of I2
    i2_value = ws['I2'].value

    # Component 1: I2 contains a formula (0.4 points)
    # This FAILS on initial (I2 is None) and PASSES on golden (I2 has formula)
    try:
        if i2_value is not None and isinstance(i2_value, str) and i2_value.strip().startswith('='):
            print(f"PASS: Component 1 — I2 contains a formula: {i2_value} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — I2 does not contain a formula, found: {repr(i2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula uses INDEX + MATCH for two-way lookup (0.3 points)
    # Must contain INDEX and at least two MATCH calls (row match + column match)
    try:
        if i2_value is not None and isinstance(i2_value, str):
            formula_upper = i2_value.upper().replace(' ', '')
            has_index = 'INDEX(' in formula_upper
            # Count MATCH occurrences — need at least 2 for two-way lookup
            match_count = formula_upper.count('MATCH(')
            if has_index and match_count >= 2:
                print(f"PASS: Component 2 — Formula uses INDEX with {match_count} MATCH calls (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Expected INDEX + 2x MATCH. INDEX={has_index}, MATCH count={match_count}")
        else:
            print(f"FAIL: Component 2 — I2 is not a formula string: {repr(i2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula is functionally correct — references correct ranges (0.3 points)
    # The correct formula: =INDEX(B2:E4,MATCH(G2,A2:A4,0),MATCH(H2,B1:E1,0))
    # We verify:
    #   - INDEX array covers the data range (B2:E4)
    #   - First MATCH looks up G2 (term) in A2:A4 (row headers) with exact match
    #   - Second MATCH looks up H2 (credit) in B1:E1 (column headers) with exact match
    try:
        if i2_value is not None and isinstance(i2_value, str):
            # Normalize: uppercase, remove spaces
            norm = i2_value.upper().replace(' ', '')

            # Check the key structural elements of the formula
            checks_passed = 0
            total_checks = 4

            # Check 1: INDEX array includes the data range B2:E4
            if 'B2:E4' in norm:
                checks_passed += 1
            else:
                print(f"  INFO: Component 3 — INDEX array range not B2:E4")

            # Check 2: MATCH for row lookup references G2 and A2:A4
            if 'G2' in norm and 'A2:A4' in norm:
                checks_passed += 1
            else:
                print(f"  INFO: Component 3 — Row MATCH missing G2 or A2:A4")

            # Check 3: MATCH for column lookup references H2 and B1:E1
            if 'H2' in norm and 'B1:E1' in norm:
                checks_passed += 1
            else:
                print(f"  INFO: Component 3 — Column MATCH missing H2 or B1:E1")

            # Check 4: Both MATCH use exact match (0)
            # Pattern: MATCH(...,0) should appear twice
            exact_match_count = len(re.findall(r'MATCH\([^)]*,0\)', norm))
            if exact_match_count >= 2:
                checks_passed += 1
            else:
                print(f"  INFO: Component 3 — Expected 2 exact-match (,0) parameters, found {exact_match_count}")

            if checks_passed == total_checks:
                print(f"PASS: Component 3 — Formula references are correct ({checks_passed}/{total_checks} sub-checks) (0.3 pts)")
                total_score += 0.3
            elif checks_passed >= 3:
                partial = 0.2
                print(f"PARTIAL: Component 3 — {checks_passed}/{total_checks} sub-checks passed ({partial} pts)")
                total_score += partial
            elif checks_passed >= 2:
                partial = 0.1
                print(f"PARTIAL: Component 3 — {checks_passed}/{total_checks} sub-checks passed ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {checks_passed}/{total_checks} sub-checks passed")
        else:
            print(f"FAIL: Component 3 — I2 is not a formula string: {repr(i2_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
