"""
Initial Setup: Two-way INDEX/MATCH/MATCH interest rate lookup
Task ID: calc_lf_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Rates ---
    ws = wb.active
    ws.title = 'Rates'

    # Header row: A1 empty, B1-E1 credit ratings
    ws.cell(row=1, column=2, value='Excellent')
    ws.cell(row=1, column=3, value='Good')
    ws.cell(row=1, column=4, value='Fair')
    ws.cell(row=1, column=5, value='Poor')

    # Data rows: term labels in column A, rates in B-E
    rate_data = [
        ['15-year', 3.25, 3.75, 4.50, 5.25],
        ['20-year', 3.50, 4.00, 4.75, 5.50],
        ['30-year', 3.75, 4.25, 5.00, 5.75],
    ]
    for r, row_data in enumerate(rate_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Lookup section: G1:I2
    ws.cell(row=1, column=7, value='Term')
    ws.cell(row=2, column=7, value='30-year')
    ws.cell(row=1, column=8, value='Credit')
    ws.cell(row=2, column=8, value='Good')
    ws.cell(row=1, column=9, value='Rate')
    # I2 intentionally left empty - the agent must enter the formula

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
