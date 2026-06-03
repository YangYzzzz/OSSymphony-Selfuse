"""
Initial Setup: Apply color scale conditional formatting to performance scores
Task ID: calc_gsd_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance"

    # Headers
    headers = ["Employee ID", "Name", "Department", "Performance Score", "Bonus%", "Manager"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 40 employee records with realistic data
    departments = ["Engineering", "Marketing", "Sales", "Finance", "HR",
                   "Operations", "Legal", "Product", "Design", "Support"]

    managers = ["Lisa Wang", "Robert Kim", "Patricia Hernandez", "James O'Brien",
                "Anita Patel", "David Nakamura", "Rachel Foster", "Thomas Wright"]

    employees = [
        ("E1001", "Sarah Chen", "Engineering", 92, 15, "Lisa Wang"),
        ("E1002", "Marcus Johnson", "Marketing", 78, 10, "Robert Kim"),
        ("E1003", "Priya Sharma", "Sales", 85, 12, "Patricia Hernandez"),
        ("E1004", "James O'Connor", "Finance", 67, 7, "James O'Brien"),
        ("E1005", "Aisha Mohammed", "HR", 91, 14, "Anita Patel"),
        ("E1006", "Carlos Rivera", "Engineering", 73, 8, "Lisa Wang"),
        ("E1007", "Emily Tanaka", "Operations", 88, 13, "David Nakamura"),
        ("E1008", "David Park", "Legal", 56, 5, "Rachel Foster"),
        ("E1009", "Fatima Al-Hassan", "Product", 95, 16, "Thomas Wright"),
        ("E1010", "Michael Brown", "Design", 62, 6, "Robert Kim"),
        ("E1011", "Sofia Petrov", "Engineering", 98, 18, "Lisa Wang"),
        ("E1012", "William Chang", "Marketing", 71, 8, "Robert Kim"),
        ("E1013", "Olivia Santos", "Sales", 83, 11, "Patricia Hernandez"),
        ("E1014", "Ahmed Khalil", "Finance", 45, 3, "James O'Brien"),
        ("E1015", "Hannah Mueller", "HR", 76, 9, "Anita Patel"),
        ("E1016", "Daniel Lee", "Operations", 89, 13, "David Nakamura"),
        ("E1017", "Grace Okafor", "Engineering", 94, 16, "Lisa Wang"),
        ("E1018", "Ryan Mitchell", "Legal", 58, 5, "Rachel Foster"),
        ("E1019", "Mei Lin", "Product", 87, 12, "Thomas Wright"),
        ("E1020", "Kevin Torres", "Design", 69, 7, "Robert Kim"),
        ("E1021", "Anna Kowalski", "Sales", 82, 11, "Patricia Hernandez"),
        ("E1022", "Jason Nguyen", "Engineering", 91, 14, "Lisa Wang"),
        ("E1023", "Rebecca Adams", "Marketing", 74, 9, "Robert Kim"),
        ("E1024", "Samuel Osei", "Finance", 63, 6, "James O'Brien"),
        ("E1025", "Laura Garcia", "HR", 86, 12, "Anita Patel"),
        ("E1026", "Nathan Brooks", "Operations", 77, 9, "David Nakamura"),
        ("E1027", "Yuki Watanabe", "Engineering", 96, 17, "Lisa Wang"),
        ("E1028", "Christopher Hall", "Legal", 52, 4, "Rachel Foster"),
        ("E1029", "Isabella Cruz", "Product", 90, 14, "Thomas Wright"),
        ("E1030", "Brian Thompson", "Design", 65, 7, "Robert Kim"),
        ("E1031", "Zara Hussain", "Sales", 81, 10, "Patricia Hernandez"),
        ("E1032", "Patrick Kelly", "Engineering", 72, 8, "Lisa Wang"),
        ("E1033", "Nicole Fischer", "Marketing", 84, 11, "Robert Kim"),
        ("E1034", "Andre Williams", "Finance", 48, 3, "James O'Brien"),
        ("E1035", "Chloe Dubois", "HR", 79, 10, "Anita Patel"),
        ("E1036", "Tyler Martin", "Operations", 93, 15, "David Nakamura"),
        ("E1037", "Leah Bergstrom", "Engineering", 68, 7, "Lisa Wang"),
        ("E1038", "Victor Popov", "Legal", 55, 4, "Rachel Foster"),
        ("E1039", "Diana Reyes", "Product", 87, 12, "Thomas Wright"),
        ("E1040", "George Wilson", "Support", 60, 6, "Anita Patel"),
    ]

    for r, emp in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp[0])  # Employee ID
        ws.cell(row=r, column=2, value=emp[1])  # Name
        ws.cell(row=r, column=3, value=emp[2])  # Department
        ws.cell(row=r, column=4, value=emp[3])  # Performance Score
        ws.cell(row=r, column=5, value=emp[4])  # Bonus%
        ws.cell(row=r, column=6, value=emp[5])  # Manager

    # Set column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 22

    # No conditional formatting -- that's the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
