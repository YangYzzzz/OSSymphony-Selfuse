"""
Initial Setup: ROI calculation task - project investment data without ROI formulas
Task ID: osworld_calc_gross_profit_sheet2_concat_013
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Projects ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Project', 'Investment', 'Returns', 'ROI %']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic project investment data (11 rows, D column empty - ROI not calculated)
    data = [
        ['Alpha Infrastructure',  250000,  312500],
        ['Beta Software Suite',   180000,  243000],
        ['Gamma Retail Expansion',320000,  384000],
        ['Delta Logistics Hub',   415000,  456500],
        ['Epsilon R&D Program',   95000,   118750],
        ['Zeta Cloud Migration',  140000,  196000],
        ['Eta Marketing Campaign', 75000,   90000],
        ['Theta Data Center',     560000,  616000],
        ['Iota Mobile Platform',  200000,  270000],
        ['Kappa Renewable Energy',480000,  614400],
        ['Lambda Healthcare App', 165000,  214500],
    ]

    for r, row_data in enumerate(data, 2):
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        ws1.cell(row=r, column=3, value=row_data[2])
        # Column D (ROI %) intentionally left empty — agent must fill with formula

    # --- Sheet2: Summary ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2 A1 intentionally left empty — agent must add portfolio ROI formula

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
