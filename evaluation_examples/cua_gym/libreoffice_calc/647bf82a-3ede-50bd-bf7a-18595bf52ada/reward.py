"""
Reward Script: Calculate ROI percentages in Sheet1 column D and create Portfolio ROI text in Sheet2 A1
Task ID: osworld_calc_gross_profit_sheet2_concat_013
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.5): ROI formulas present in Sheet1 D2:D12 (11 rows)
  Component 2 (0.3): ROI formulas follow correct pattern (C-B)/B*100
  Component 3 (0.2): Sheet2 A1 contains formula with 'Portfolio ROI:' and AVERAGE of Sheet1 D column
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_013'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — gate check
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist — precondition gate
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    if 'Sheet2' not in wb.sheetnames:
        print("CRITICAL: Sheet2 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']

    # Component 1: ROI formulas present in Sheet1 D2:D12 (0.5 points)
    # Task requires column D to have ROI formulas (was empty in initial state)
    try:
        roi_cells_with_formula = 0
        roi_cells_checked = 0
        for row in range(2, 13):  # rows 2 through 12 (11 projects)
            cell = ws1.cell(row=row, column=4)  # Column D
            roi_cells_checked += 1
            val = cell.value
            if val is not None and isinstance(val, str) and val.startswith('='):
                roi_cells_with_formula += 1

        print(f"INFO: {roi_cells_with_formula}/{roi_cells_checked} ROI cells have formulas")

        if roi_cells_with_formula == 11:
            print(f"PASS: Component 1 — All 11 ROI cells (D2:D12) have formulas (0.5 pts)")
            total_score += 0.5
        elif roi_cells_with_formula >= 6:
            # Partial: more than half the cells have formulas
            partial = round(0.5 * roi_cells_with_formula / 11, 2)
            print(f"PARTIAL: Component 1 — {roi_cells_with_formula}/11 ROI cells have formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {roi_cells_with_formula}/11 ROI cells have formulas (expected 11)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: ROI formula correctness — pattern (C-B)/B*100 (0.3 points)
    # Each formula should follow: =(Cx-Bx)/Bx*100
    try:
        correct_formula_count = 0
        for row in range(2, 13):
            cell = ws1.cell(row=row, column=4)
            val = cell.value
            if val is not None and isinstance(val, str) and val.startswith('='):
                # Normalize formula: strip spaces, uppercase
                normalized = val.upper().replace(' ', '')
                # Pattern: =(Cx-Bx)/Bx*100 where x matches the row number
                # Accept variations like =(C2-B2)/B2*100 or =(RETURNS-INVESTMENT)/INVESTMENT*100
                # Check for the structural pattern: subtraction, division, multiply by 100
                roi_pattern = re.search(
                    r'=\(C(\d+)-B(\d+)\)/B(\d+)\*100',
                    normalized
                )
                if roi_pattern:
                    r1, r2, r3 = roi_pattern.group(1), roi_pattern.group(2), roi_pattern.group(3)
                    # All row references should match the current row
                    if r1 == r2 == r3 == str(row):
                        correct_formula_count += 1
                    else:
                        print(f"  WARN: Row {row} formula has mismatched row refs: {val}")
                else:
                    # Also accept (RETURNS-INVESTMENT)/INVESTMENT*100 style or other valid expressions
                    # Check for at minimum: subtraction and division and *100
                    if '-' in normalized and '/' in normalized and '*100' in normalized:
                        correct_formula_count += 1
                        print(f"  INFO: Row {row} uses acceptable alternative ROI formula: {val}")
                    else:
                        print(f"  FAIL: Row {row} formula does not follow ROI pattern: {val}")

        print(f"INFO: {correct_formula_count}/11 ROI formulas have correct pattern")

        if correct_formula_count == 11:
            print(f"PASS: Component 2 — All 11 ROI formulas follow correct (C-B)/B*100 pattern (0.3 pts)")
            total_score += 0.3
        elif correct_formula_count >= 6:
            partial = round(0.3 * correct_formula_count / 11, 2)
            print(f"PARTIAL: Component 2 — {correct_formula_count}/11 ROI formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {correct_formula_count}/11 ROI formulas correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet2 A1 contains Portfolio ROI formula (0.2 points)
    # Must contain 'Portfolio ROI:' text and reference AVERAGE of Sheet1 D column
    try:
        a1_val = ws2['A1'].value
        print(f"INFO: Sheet2 A1 = {repr(a1_val)}")

        if a1_val is None:
            print("FAIL: Component 3 — Sheet2 A1 is empty (expected Portfolio ROI formula)")
        elif isinstance(a1_val, str):
            # Check if it's a formula
            if a1_val.startswith('='):
                normalized = a1_val.upper().replace(' ', '')
                # Must contain 'PORTFOLIO ROI:' text string
                has_portfolio_roi = 'PORTFOLIOROI:' in normalized.replace(' ', '') or \
                                    '"PORTFOLIOROI:' in normalized.replace(' ', '') or \
                                    'PORTFOLIO ROI:' in a1_val.upper()
                # Must reference AVERAGE of Sheet1 D column
                has_average = 'AVERAGE(' in normalized
                has_sheet1_d = 'SHEET1!D' in normalized
                # Must concatenate with percentage
                has_percent = '"%"' in a1_val or "'%'" in a1_val or '&"%"' in normalized

                if has_portfolio_roi and has_average and has_sheet1_d:
                    print(f"PASS: Component 3 — Sheet2 A1 has Portfolio ROI formula referencing AVERAGE(Sheet1!D) (0.2 pts)")
                    total_score += 0.2
                elif has_average and has_sheet1_d:
                    print(f"PARTIAL: Component 3 — Sheet2 A1 has AVERAGE(Sheet1!D) but missing 'Portfolio ROI:' text (0.1 pts)")
                    total_score += 0.1
                else:
                    print(f"FAIL: Component 3 — Sheet2 A1 formula missing required elements. "
                          f"portfolio_roi={has_portfolio_roi}, average={has_average}, sheet1_d={has_sheet1_d}")
            else:
                # Static text value (not a formula) — check if it at least has the right content
                if 'portfolio roi:' in a1_val.lower():
                    print(f"PARTIAL: Component 3 — Sheet2 A1 has Portfolio ROI text but not as formula (0.05 pts)")
                    total_score += 0.05
                else:
                    print(f"FAIL: Component 3 — Sheet2 A1 is not empty but doesn't contain expected content: {repr(a1_val)}")
        else:
            print(f"FAIL: Component 3 — Sheet2 A1 has unexpected type {type(a1_val)}: {repr(a1_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
