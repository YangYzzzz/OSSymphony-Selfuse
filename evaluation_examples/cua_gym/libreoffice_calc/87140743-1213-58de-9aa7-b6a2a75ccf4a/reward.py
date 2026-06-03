"""
Reward Script: Apply Standard Filter to Show Only Software and Hardware Assets
Task ID: calc_dop_filter_standard_071
Domain: libreoffice_calc

Task: Use the Standard Filter (Data -> Standard Filter) with the criteria range G1:G3
to show only asset records where Category is 'Software' OR Category is 'Hardware'.

Scoring Rubric:
  Component 1: Correct rows are hidden — rows with Furniture/Vehicles/Other categories
               are all hidden (0.6 pts)
  Component 2: All visible data rows have Category = Software or Hardware, and
               the count matches exactly 62 visible data rows (0.4 pts)
  Total: 1.0

Note: Criteria range G1:G3 (Category/Software/Hardware) is a precondition that exists
in both initial and golden files; it is NOT scored but used as a data integrity gate.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_filter_standard_071'
SHEET_NAME = 'AssetRegister'

# Expected counts from context:
# Software: 28, Hardware: 34 => visible data rows = 62
# Furniture: 15, Vehicles: 12, Other: 10 => hidden rows = 37
EXPECTED_HIDDEN_COUNT = 37
EXPECTED_VISIBLE_DATA_COUNT = 62
HIDDEN_CATEGORIES = {'Furniture', 'Vehicles', 'Other'}
VISIBLE_CATEGORIES = {'Software', 'Hardware'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Precondition gate: criteria range G1:G3 must still be intact (not a scored component)
    g1 = ws.cell(row=1, column=7).value
    g2 = ws.cell(row=2, column=7).value
    g3 = ws.cell(row=3, column=7).value
    if str(g1).strip() != 'Category' or str(g2).strip() != 'Software' or str(g3).strip() != 'Hardware':
        print(f"CRITICAL: Criteria range G1:G3 has been modified: G1={g1!r}, G2={g2!r}, G3={g3!r}")
        print("REWARD: 0.0")
        return 0.0
    else:
        print(f"PRECONDITION OK: Criteria range G1:G3 intact: G1={g1!r}, G2={g2!r}, G3={g3!r}")

    # Component 1: Rows with Furniture/Vehicles/Other categories are hidden (0.6 points)
    # The filter should hide exactly 37 rows — all rows with Category NOT in {Software, Hardware}.
    # Initial file: 0 hidden rows => FAILS this check.
    # Golden file: 37 rows hidden (Furniture:15, Other:10, Vehicles:12) => PASSES.
    try:
        hidden_rows = sorted([r for r, rd in ws.row_dimensions.items() if rd.hidden])
        hidden_count = len(hidden_rows)

        # Check all hidden rows belong to the non-target categories
        wrong_hidden = []
        for r in hidden_rows:
            cat = ws.cell(row=r, column=3).value
            if cat in VISIBLE_CATEGORIES:
                wrong_hidden.append((r, cat))

        # Check all non-target category rows are actually hidden
        missed_rows = []
        for r in range(2, 101):
            cat = ws.cell(row=r, column=3).value
            if cat in HIDDEN_CATEGORIES and r not in set(hidden_rows):
                missed_rows.append((r, cat))

        if hidden_count == EXPECTED_HIDDEN_COUNT and len(wrong_hidden) == 0 and len(missed_rows) == 0:
            print(f"PASS: Component 1 — {hidden_count} rows correctly hidden (all Furniture/Vehicles/Other) (0.6 pts)")
            total_score += 0.6
        elif hidden_count > 0 and len(wrong_hidden) == 0 and hidden_count < EXPECTED_HIDDEN_COUNT:
            # Partial: some non-target rows hidden but not all
            partial = round(0.6 * hidden_count / EXPECTED_HIDDEN_COUNT, 2)
            print(f"PARTIAL: Component 1 — {hidden_count}/{EXPECTED_HIDDEN_COUNT} non-target rows hidden, no wrong hides ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected {EXPECTED_HIDDEN_COUNT} hidden rows, found {hidden_count}. "
                  f"Wrong hidden: {wrong_hidden[:3]}, Missed: {missed_rows[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All visible data rows have Category = Software or Hardware,
    # and the exact count matches 62 (0.4 points)
    # Initial file: 99 visible data rows, 37 have wrong category => FAILS.
    # Golden file: 62 visible data rows, all in {Software, Hardware} => PASSES.
    try:
        hidden_rows_set = set(r for r, rd in ws.row_dimensions.items() if rd.hidden)
        visible_data_rows = [r for r in range(2, 101) if r not in hidden_rows_set]
        visible_count = len(visible_data_rows)

        wrong_visible = []
        for r in visible_data_rows:
            cat = ws.cell(row=r, column=3).value
            if cat not in VISIBLE_CATEGORIES:
                wrong_visible.append((r, cat))

        if len(wrong_visible) == 0 and visible_count == EXPECTED_VISIBLE_DATA_COUNT:
            print(f"PASS: Component 2 — {visible_count} visible data rows, all Category in {{Software, Hardware}} (0.4 pts)")
            total_score += 0.4
        elif len(wrong_visible) == 0 and visible_count > 0:
            print(f"PARTIAL: Component 2 — {visible_count} visible rows, all have correct categories but expected {EXPECTED_VISIBLE_DATA_COUNT} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — {len(wrong_visible)} visible rows have wrong category: {wrong_visible[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
