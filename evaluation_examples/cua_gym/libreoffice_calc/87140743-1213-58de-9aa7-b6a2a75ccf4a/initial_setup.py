"""
Initial Setup: Standard Filter - AssetRegister with Software/Hardware criteria
Task ID: calc_dop_filter_standard_071
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_filter_standard_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AssetRegister'

    # --- Headers in row 1 ---
    headers = ['Asset ID', 'Asset Name', 'Category', 'Purchase Date', 'Value', 'Location']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Category criteria range: G1:G3 ---
    ws['G1'] = 'Category'
    ws['G2'] = 'Software'
    ws['G3'] = 'Hardware'
    ws['G1'].font = Font(bold=True)

    # --- Asset data: 99 rows (rows 2–100) ---
    # Distribution: Software(28), Hardware(34), Furniture(15), Vehicles(12), Other(10) = 99
    software_assets = [
        ('Adobe Acrobat Pro', 'Software', '2023-03-10', 1250.00, 'HQ - IT Dept'),
        ('Microsoft Office 365', 'Software', '2023-01-15', 2800.00, 'HQ - Finance'),
        ('AutoCAD 2024', 'Software', '2022-11-20', 4500.00, 'Branch - Engineering'),
        ('Slack Business', 'Software', '2023-05-01', 960.00, 'HQ - Operations'),
        ('Zoom Pro License', 'Software', '2023-02-14', 720.00, 'Remote'),
        ('Salesforce CRM', 'Software', '2022-09-30', 6200.00, 'HQ - Sales'),
        ('GitHub Enterprise', 'Software', '2023-06-01', 3150.00, 'HQ - IT Dept'),
        ('Jira Software', 'Software', '2022-12-05', 1800.00, 'HQ - IT Dept'),
        ('Tableau Desktop', 'Software', '2023-04-18', 2100.00, 'HQ - Analytics'),
        ('QuickBooks Enterprise', 'Software', '2022-08-22', 3400.00, 'HQ - Finance'),
        ('Adobe Creative Cloud', 'Software', '2023-07-10', 1640.00, 'HQ - Marketing'),
        ('SAP ERP Module', 'Software', '2022-06-15', 12000.00, 'HQ - Operations'),
        ('Dropbox Business', 'Software', '2023-01-28', 540.00, 'Remote'),
        ('Asana Premium', 'Software', '2023-03-22', 480.00, 'HQ - Projects'),
        ('AWS Enterprise Support', 'Software', '2022-10-01', 8500.00, 'Cloud'),
        ('Figma Organization', 'Software', '2023-05-15', 1200.00, 'HQ - Design'),
        ('Power BI Pro', 'Software', '2023-02-01', 900.00, 'HQ - Analytics'),
        ('Visual Studio Enterprise', 'Software', '2022-07-20', 3600.00, 'HQ - IT Dept'),
        ('Docusign Business Pro', 'Software', '2023-04-05', 750.00, 'HQ - Legal'),
        ('Zendesk Suite', 'Software', '2022-11-10', 2400.00, 'HQ - Support'),
        ('1Password Teams', 'Software', '2023-06-20', 480.00, 'HQ - IT Dept'),
        ('Notion Team Plan', 'Software', '2023-01-10', 360.00, 'Remote'),
        ('HubSpot Marketing', 'Software', '2022-09-05', 4800.00, 'HQ - Marketing'),
        ('Confluence Data Center', 'Software', '2023-03-01', 2700.00, 'HQ - IT Dept'),
        ('LogMeIn Rescue', 'Software', '2022-08-15', 1320.00, 'HQ - Support'),
        ('Datadog Pro', 'Software', '2023-02-20', 3900.00, 'Cloud'),
        ('Sketch Teams', 'Software', '2023-05-05', 720.00, 'HQ - Design'),
        ('Monday.com Enterprise', 'Software', '2022-12-20', 1800.00, 'HQ - Projects'),
    ]

    hardware_assets = [
        ('Dell Latitude 5540', 'Hardware', '2023-02-10', 1450.00, 'HQ - IT Dept'),
        ('HP ProBook 450 G10', 'Hardware', '2023-01-20', 1280.00, 'Branch - Sales'),
        ('Lenovo ThinkPad X1', 'Hardware', '2022-11-15', 1850.00, 'HQ - Management'),
        ('Apple MacBook Pro 14"', 'Hardware', '2023-04-05', 2399.00, 'HQ - Design'),
        ('Cisco Catalyst 9200', 'Hardware', '2022-09-10', 3600.00, 'HQ - Network'),
        ('HP LaserJet Pro 4001n', 'Hardware', '2023-03-15', 540.00, 'HQ - Admin'),
        ('Dell OptiPlex 7010', 'Hardware', '2023-05-20', 980.00, 'Branch - Finance'),
        ('Polycom Studio X30', 'Hardware', '2022-12-10', 1200.00, 'HQ - Conf Room A'),
        ('Samsung 27" Monitor', 'Hardware', '2023-01-05', 420.00, 'HQ - IT Dept'),
        ('Logitech MX Master 3', 'Hardware', '2023-06-10', 95.00, 'HQ - IT Dept'),
        ('APC Smart-UPS 1500', 'Hardware', '2022-10-20', 680.00, 'HQ - Server Room'),
        ('Cisco RV340 Router', 'Hardware', '2023-02-25', 440.00, 'Branch - Office'),
        ('HP Z4 Workstation', 'Hardware', '2022-08-05', 3200.00, 'HQ - Engineering'),
        ('LG UltraWide 34"', 'Hardware', '2023-03-30', 680.00, 'HQ - Analytics'),
        ('Dell PowerEdge R750', 'Hardware', '2022-07-15', 8500.00, 'HQ - Server Room'),
        ('Epson WorkForce WF-7840', 'Hardware', '2023-04-20', 380.00, 'Branch - Admin'),
        ('Synology DS923+', 'Hardware', '2022-11-25', 890.00, 'HQ - IT Dept'),
        ('Palo Alto PA-220', 'Hardware', '2023-01-30', 2100.00, 'HQ - Network'),
        ('Apple iPad Pro 12.9"', 'Hardware', '2023-05-10', 1299.00, 'HQ - Field Sales'),
        ('Jabra Evolve2 85', 'Hardware', '2023-02-15', 280.00, 'Remote'),
        ('Wacom Cintiq 16', 'Hardware', '2022-09-25', 650.00, 'HQ - Design'),
        ('HP DeskJet 4120e', 'Hardware', '2023-06-05', 150.00, 'Branch - HR'),
        ('Cisco IP Phone 7945', 'Hardware', '2022-12-20', 320.00, 'HQ - Reception'),
        ('Dell Dock WD22TB4', 'Hardware', '2023-03-08', 260.00, 'HQ - IT Dept'),
        ('Raspberry Pi 4 8GB', 'Hardware', '2023-04-15', 85.00, 'HQ - Lab'),
        ('Barco ClickShare CX-30', 'Hardware', '2022-08-30', 1600.00, 'HQ - Conf Room B'),
        ('Brother MFC-L9570CDW', 'Hardware', '2023-05-25', 920.00, 'HQ - Print Room'),
        ('Netgear ReadyNAS 526X', 'Hardware', '2022-10-10', 1100.00, 'Branch - IT'),
        ('Plantronics CS540', 'Hardware', '2023-01-18', 175.00, 'HQ - Call Centre'),
        ('HP EliteDisplay E27', 'Hardware', '2023-06-15', 340.00, 'HQ - Finance'),
        ('Cisco SG350 Switch', 'Hardware', '2022-11-05', 580.00, 'Branch - Network'),
        ('Logitech Rally Camera', 'Hardware', '2023-02-28', 990.00, 'HQ - Conf Room C'),
        ('Apple Mac mini M2', 'Hardware', '2023-04-12', 1099.00, 'HQ - Dev Lab'),
        ('Acronis Backup Appliance', 'Hardware', '2022-07-28', 2800.00, 'HQ - IT Dept'),
    ]

    furniture_assets = [
        ('Herman Miller Aeron Chair', 'Furniture', '2022-06-01', 1395.00, 'HQ - Exec Suite'),
        ('IKEA GALANT Desk 160x80', 'Furniture', '2023-01-10', 320.00, 'Branch - Open Plan'),
        ('Steelcase Leap Chair V2', 'Furniture', '2022-09-15', 1180.00, 'HQ - Management'),
        ('Knoll Dividends Lateral File', 'Furniture', '2023-02-20', 890.00, 'HQ - Admin'),
        ('Humanscale Float Table', 'Furniture', '2022-11-01', 2100.00, 'HQ - Conf Room A'),
        ('HON 94000 Series Workstation', 'Furniture', '2023-03-15', 1560.00, 'Branch - Finance'),
        ('Vari Electric Stand Desk', 'Furniture', '2023-04-10', 695.00, 'HQ - IT Dept'),
        ('Ikea Bekant Sit/Stand Desk', 'Furniture', '2022-08-20', 450.00, 'Remote'),
        ('Haworth Fern Chair', 'Furniture', '2023-05-01', 1620.00, 'HQ - Design'),
        ('Global Total Office Wardrobe', 'Furniture', '2022-10-05', 780.00, 'HQ - Breakroom'),
        ('Safco Tuff Stor Cabinet', 'Furniture', '2023-01-25', 360.00, 'Branch - Storage'),
        ('Lorell Essentials Bookcase', 'Furniture', '2022-12-10', 240.00, 'HQ - Library'),
        ('Teknion Leverage Panel', 'Furniture', '2023-02-05', 520.00, 'HQ - Open Floor'),
        ('Humanscale M8 Monitor Arm', 'Furniture', '2023-03-20', 310.00, 'HQ - IT Dept'),
        ('Mayline Conference Table 12ft', 'Furniture', '2022-07-10', 2800.00, 'HQ - Conf Room B'),
    ]

    vehicles_assets = [
        ('Toyota Camry 2022', 'Vehicles', '2022-05-15', 28500.00, 'Fleet - Sydney'),
        ('Ford Transit Van 2021', 'Vehicles', '2021-11-20', 42000.00, 'Fleet - Melbourne'),
        ('Honda CR-V 2023', 'Vehicles', '2023-01-08', 36800.00, 'Fleet - Brisbane'),
        ('Mitsubishi Outlander 2022', 'Vehicles', '2022-08-22', 39500.00, 'Fleet - Perth'),
        ('Toyota HiLux 2022', 'Vehicles', '2022-06-30', 48000.00, 'Fleet - Adelaide'),
        ('Hyundai Tucson 2023', 'Vehicles', '2023-03-05', 34900.00, 'Fleet - Canberra'),
        ('Isuzu D-Max 2022', 'Vehicles', '2022-09-18', 52000.00, 'Fleet - Darwin'),
        ('Ford Ranger 2023', 'Vehicles', '2023-02-14', 56000.00, 'Fleet - Hobart'),
        ('Nissan Navara 2022', 'Vehicles', '2022-11-01', 47000.00, 'Fleet - Sydney'),
        ('Kia Sportage 2023', 'Vehicles', '2023-04-20', 33500.00, 'Fleet - Melbourne'),
        ('Toyota LandCruiser 200', 'Vehicles', '2022-07-12', 88000.00, 'Fleet - Remote Ops'),
        ('Volkswagen Transporter', 'Vehicles', '2022-10-05', 58000.00, 'Fleet - Brisbane'),
    ]

    other_assets = [
        ('Nikon D6 Camera Kit', 'Other', '2022-08-10', 6500.00, 'HQ - Marketing'),
        ('Canon EOS R5 Mirrorless', 'Other', '2023-02-01', 4800.00, 'HQ - Media'),
        ('DJI Mavic 3 Pro Drone', 'Other', '2023-05-10', 2200.00, 'HQ - Surveys'),
        ('Sony A7 IV Camera', 'Other', '2022-11-15', 3100.00, 'HQ - Marketing'),
        ('Leica M11 Camera', 'Other', '2023-01-20', 8900.00, 'HQ - Executive'),
        ('Sony WH-1000XM5 Headset', 'Other', '2023-03-05', 380.00, 'HQ - Podcasting'),
        ('Rode NT1 Microphone Kit', 'Other', '2022-09-20', 420.00, 'HQ - Studio'),
        ('Elgato Stream Deck XL', 'Other', '2023-04-15', 250.00, 'HQ - Media'),
        ('Pelican 1650 Case', 'Other', '2022-12-10', 320.00, 'Field - Equipment'),
        ('Makita Power Drill Set', 'Other', '2023-06-01', 180.00, 'Facilities'),
    ]

    # Interleave all assets to create realistic mixed data
    all_assets = []
    # Add them in a mixed order
    indices = list(range(28))   # software indices
    soft_idx = 0
    hard_idx = 0
    furn_idx = 0
    vehi_idx = 0
    othr_idx = 0

    # Create the 99-row sequence in a specific order that mixes categories
    sequence = (
        ['S'] * 28 +
        ['H'] * 34 +
        ['F'] * 15 +
        ['V'] * 12 +
        ['O'] * 10
    )
    # Shuffle with a fixed seed for reproducibility
    import random
    random.seed(42)
    random.shuffle(sequence)

    asset_id_counter = 1000
    for cat_code in sequence:
        if cat_code == 'S' and soft_idx < len(software_assets):
            row = software_assets[soft_idx]
            soft_idx += 1
        elif cat_code == 'H' and hard_idx < len(hardware_assets):
            row = hardware_assets[hard_idx]
            hard_idx += 1
        elif cat_code == 'F' and furn_idx < len(furniture_assets):
            row = furniture_assets[furn_idx]
            furn_idx += 1
        elif cat_code == 'V' and vehi_idx < len(vehicles_assets):
            row = vehicles_assets[vehi_idx]
            vehi_idx += 1
        elif cat_code == 'O' and othr_idx < len(other_assets):
            row = other_assets[othr_idx]
            othr_idx += 1
        else:
            continue

        all_assets.append(row)
        asset_id_counter += 1

    # Write data rows
    for r, (name, cat, date, value, location) in enumerate(all_assets, 2):
        asset_id = f'AST-{1000 + r - 2:04d}'
        ws.cell(row=r, column=1, value=asset_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=date)
        ws.cell(row=r, column=5, value=value)
        ws.cell(row=r, column=6, value=location)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 26
    ws.column_dimensions['G'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: AssetRegister')
    print(f'  Data rows: {len(all_assets)} (rows 2–{len(all_assets)+1})')
    print(f'  Criteria range: G1:G3 (Category / Software / Hardware)')


create_initial()
