"""
Reward Script: Calculate Net Promoter Score from customer survey data
Task ID: calc_sales_nps_analysis_064
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): IFS formula in NPSSurvey C2:C501 classifying scores into Promoter/Passive/Detractor
  - Component 2 (0.25): COUNTIFS formulas in NPSResults B2:B4 counting each NPS category
  - Component 3 (0.20): Percentage formulas in NPSResults C2:C4 with percentage number format
  - Component 4 (0.15): NPS Score formula in NPSResults B6 (=C2-C4)
  - Component 5 (0.10): Stacked bar chart on NPSResults sheet
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_nps_analysis_064'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist as precondition
    if 'NPSSurvey' not in wb.sheetnames or 'NPSResults' not in wb.sheetnames:
        print("CRITICAL: Required sheets 'NPSSurvey' and/or 'NPSResults' not found")
        print("REWARD: 0.0")
        return 0.0

    ws_survey = wb['NPSSurvey']
    ws_results = wb['NPSResults']

    # Component 1: IFS formula in NPSSurvey C2:C501 classifying scores (0.30 points)
    # The task requires classifying respondents: Promoters (9-10), Passives (7-8), Detractors (0-6)
    # This FAILS on initial (all None) and PASSES on golden (all have IFS formulas)
    try:
        formula_count = 0
        valid_formula_count = 0
        total_rows = 500  # rows 2 to 501

        for row in range(2, 502):
            val = ws_survey.cell(row=row, column=3).value
            if val is not None and isinstance(val, str):
                formula_count += 1
                # Check that the formula classifies correctly: must reference B column and
                # use IFS (or IF) to categorize as Promoter, Passive, Detractor
                val_upper = val.upper()
                if ('IFS(' in val_upper or 'IF(' in val_upper) and \
                   'PROMOTER' in val_upper and \
                   'PASSIVE' in val_upper and \
                   'DETRACTOR' in val_upper:
                    valid_formula_count += 1

        if valid_formula_count == total_rows:
            print(f"PASS: Component 1 — All {total_rows} cells in NPSSurvey C2:C501 have valid IFS classification formulas (0.30 pts)")
            total_score += 0.30
        elif valid_formula_count >= total_rows * 0.9:
            # Partial credit: at least 90% of rows have the formula
            partial = 0.15
            print(f"PARTIAL: Component 1 — {valid_formula_count}/{total_rows} cells have valid IFS formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {valid_formula_count}/{total_rows} cells have valid IFS classification formulas (expected {total_rows})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: COUNTIFS formulas in NPSResults B2:B4 counting each NPS category (0.25 points)
    # B2 counts Promoters (scores >= 9), B3 counts Passives (7-8), B4 counts Detractors (<= 6)
    # This FAILS on initial (all None) and PASSES on golden (all have COUNTIFS formulas)
    try:
        b2 = ws_results.cell(row=2, column=2).value  # Promoters count
        b3 = ws_results.cell(row=3, column=2).value  # Passives count
        b4 = ws_results.cell(row=4, column=2).value  # Detractors count

        b2_ok = b2 is not None and isinstance(b2, str) and 'COUNTIFS' in b2.upper()
        b3_ok = b3 is not None and isinstance(b3, str) and 'COUNTIFS' in b3.upper()
        b4_ok = b4 is not None and isinstance(b4, str) and 'COUNTIFS' in b4.upper()

        passing = sum([b2_ok, b3_ok, b4_ok])
        if passing == 3:
            print(f"PASS: Component 2 — COUNTIFS formulas present in NPSResults B2:B4 (B2={repr(b2[:30])}, B3={repr(b3[:30])}, B4={repr(b4[:30])}) (0.25 pts)")
            total_score += 0.25
        elif passing > 0:
            partial = round(0.25 * passing / 3, 4)
            print(f"PARTIAL: Component 2 — {passing}/3 COUNTIFS formulas present in NPSResults B2:B4 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No COUNTIFS formulas found in NPSResults B2:B4 (B2={repr(b2)}, B3={repr(b3)}, B4={repr(b4)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage formulas in NPSResults C2:C4 with percentage number format (0.20 points)
    # C2:C4 should have =B2/SUM($B$2:$B$4) style formulas and be formatted as percentage
    # This FAILS on initial (all None) and PASSES on golden
    try:
        c2 = ws_results.cell(row=2, column=3).value
        c3 = ws_results.cell(row=3, column=3).value
        c4 = ws_results.cell(row=4, column=3).value
        c2_fmt = ws_results.cell(row=2, column=3).number_format
        c3_fmt = ws_results.cell(row=3, column=3).number_format
        c4_fmt = ws_results.cell(row=4, column=3).number_format

        # Formula check: should reference B column cells and SUM
        def is_pct_formula(val):
            if val is None or not isinstance(val, str):
                return False
            v = val.upper().replace(' ', '')
            return v.startswith('=B') and 'SUM' in v

        # Format check: percentage format
        def is_pct_format(fmt):
            return fmt is not None and '%' in str(fmt)

        c2_ok = is_pct_formula(c2) and is_pct_format(c2_fmt)
        c3_ok = is_pct_formula(c3) and is_pct_format(c3_fmt)
        c4_ok = is_pct_formula(c4) and is_pct_format(c4_fmt)

        passing = sum([c2_ok, c3_ok, c4_ok])
        if passing == 3:
            print(f"PASS: Component 3 — Percentage formulas with % format in NPSResults C2:C4 (0.20 pts)")
            total_score += 0.20
        elif passing > 0:
            partial = round(0.20 * passing / 3, 4)
            print(f"PARTIAL: Component 3 — {passing}/3 percentage formulas with % format in NPSResults C2:C4 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No percentage formulas with % format found in NPSResults C2:C4 (C2={repr(c2)}, fmt={repr(c2_fmt)})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: NPS Score formula in NPSResults B6 (=C2-C4) (0.15 points)
    # B6 should contain =C2-C4 (Promoter% - Detractor%) formatted as percentage
    # This FAILS on initial (None) and PASSES on golden
    try:
        b6 = ws_results.cell(row=6, column=2).value
        b6_fmt = ws_results.cell(row=6, column=2).number_format

        # Formula check: B6 should be =C2-C4
        def is_nps_formula(val):
            if val is None or not isinstance(val, str):
                return False
            v = val.upper().replace(' ', '')
            return v in ('=C2-C4',) or ('C2' in v and 'C4' in v and '-' in v)

        b6_formula_ok = is_nps_formula(b6)
        b6_format_ok = b6_fmt is not None and '%' in str(b6_fmt)

        if b6_formula_ok and b6_format_ok:
            print(f"PASS: Component 4 — NPS formula in NPSResults B6={repr(b6)}, format={repr(b6_fmt)} (0.15 pts)")
            total_score += 0.15
        elif b6_formula_ok:
            # Formula is correct but format might not be percentage
            print(f"PARTIAL: Component 4 — NPS formula in B6 correct (B6={repr(b6)}) but format is {repr(b6_fmt)}, not percentage (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — NPS formula missing or incorrect in NPSResults B6 (B6={repr(b6)}, expected =C2-C4 or equivalent)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Stacked bar chart on NPSResults sheet (0.10 points)
    # Task requires "a stacked bar chart showing the distribution of NPS categories"
    # This FAILS on initial (no charts) and PASSES on golden (has 1 stacked bar chart)
    try:
        charts = ws_results._charts
        if len(charts) >= 1:
            # Check if any chart is a stacked bar/column chart
            stacked_found = False
            for chart in charts:
                chart_type = type(chart).__name__
                grouping = getattr(chart, 'grouping', None)
                if 'Bar' in chart_type and grouping in ('stacked', 'percentStacked'):
                    stacked_found = True
                    print(f"PASS: Component 5 — Stacked bar chart found on NPSResults (type={chart_type}, grouping={grouping}) (0.10 pts)")
                    total_score += 0.10
                    break
            if not stacked_found:
                # A chart exists but may not be stacked
                print(f"PARTIAL: Component 5 — Chart found on NPSResults but not stacked bar (type={type(charts[0]).__name__}, grouping={getattr(charts[0], 'grouping', 'N/A')}) (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No chart found on NPSResults sheet (expected a stacked bar chart)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
