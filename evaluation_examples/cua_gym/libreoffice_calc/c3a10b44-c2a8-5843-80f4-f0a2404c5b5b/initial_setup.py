"""
Initial Setup: NPS Survey data - pre-task state
Task ID: calc_sales_nps_analysis_064
Domain: libreoffice_calc
"""

import random
import datetime
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_nps_analysis_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

# Realistic customer segments
SEGMENTS = ['Enterprise', 'SMB', 'Startup', 'Consumer', 'Government', 'Education']

# Customer first/last names for realistic IDs
FIRST_NAMES = [
    'Sarah', 'Marcus', 'Emily', 'James', 'Olivia', 'Daniel', 'Sophia', 'Michael',
    'Isabella', 'William', 'Mia', 'David', 'Charlotte', 'Joseph', 'Amelia', 'Samuel',
    'Harper', 'Benjamin', 'Evelyn', 'Lucas', 'Abigail', 'Henry', 'Ella', 'Alexander',
    'Avery', 'Jackson', 'Sofia', 'Sebastian', 'Luna', 'Matthew', 'Grace', 'Aiden',
    'Chloe', 'Owen', 'Penelope', 'Liam', 'Victoria', 'Noah', 'Riley', 'Ethan'
]

def generate_score():
    """Generate a realistic NPS score distribution:
    ~30% Detractors (0-6), ~20% Passives (7-8), ~50% Promoters (9-10)
    """
    r = random.random()
    if r < 0.30:
        return random.randint(0, 6)
    elif r < 0.50:
        return random.randint(7, 8)
    else:
        return random.randint(9, 10)

def generate_customer_id(index):
    """Generate a customer ID like CUST-001"""
    return f'CUST-{index:04d}'

def generate_date():
    """Generate a date between 2024-01-01 and 2025-06-30"""
    start = datetime.date(2024, 1, 1)
    end = datetime.date(2025, 6, 30)
    delta = (end - start).days
    return start + datetime.timedelta(days=random.randint(0, delta))


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: NPSSurvey ---
    ws1 = wb.active
    ws1.title = 'NPSSurvey'

    # Headers
    headers = ['Customer ID', 'Score', 'Category', 'Segment', 'Date']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # 500 rows of survey data
    for i in range(1, 501):
        customer_id = generate_customer_id(i)
        score = generate_score()
        category = ''          # EMPTY — to be filled by agent
        segment = random.choice(SEGMENTS)
        date = generate_date()

        ws1.cell(row=i + 1, column=1, value=customer_id)
        ws1.cell(row=i + 1, column=2, value=score)
        ws1.cell(row=i + 1, column=3, value=category)   # empty
        ws1.cell(row=i + 1, column=4, value=segment)
        ws1.cell(row=i + 1, column=5, value=date.strftime('%Y-%m-%d'))

    # --- Sheet 2: NPSResults ---
    ws2 = wb.create_sheet('NPSResults')

    # Column headers
    ws2['B1'] = 'Count'
    ws2['C1'] = 'Percentage'

    # Row labels
    ws2['A2'] = 'Promoters (9-10)'
    ws2['A3'] = 'Passives (7-8)'
    ws2['A4'] = 'Detractors (0-6)'

    # NPS label
    ws2['A6'] = 'NPS Score'

    # All value cells (B2:B4, C2:C4, B6) are intentionally left empty
    # They will be filled with formulas by the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
