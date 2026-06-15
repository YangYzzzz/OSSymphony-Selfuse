"""
Initial Setup: Sort logistics shipment data by dispatch date and create line chart
Task ID: osworld_calc_sort_date_chart_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sort_date_chart_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Shipments ---
    ws = wb.active
    ws.title = 'Shipments'

    # Headers
    headers = ['Shipment ID', 'Dispatch Date', 'Carrier', 'Destination', 'Weight kg']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic shipment data in RANDOM order (NOT sorted by date)
    # Dates deliberately shuffled so agent must sort them
    data = [
        ['SHP-20140', date(2025, 3, 14), 'FedEx',   'Berlin',       23.5],
        ['SHP-20101', date(2025, 2, 28), 'UPS',      'Tokyo',        55.0],
        ['SHP-20115', date(2025, 3, 7),  'DHL',      'Sydney',       12.3],
        ['SHP-20128', date(2025, 3, 11), 'FedEx',    'New York',     88.7],
        ['SHP-20093', date(2025, 2, 20), 'Aramex',   'Dubai',        34.2],
        ['SHP-20107', date(2025, 3, 3),  'TNT',      'Paris',        47.6],
        ['SHP-20122', date(2025, 3, 9),  'UPS',      'Singapore',    19.1],
        ['SHP-20135', date(2025, 3, 13), 'DHL',      'London',       62.4],
        ['SHP-20088', date(2025, 2, 17), 'FedEx',    'Toronto',      30.8],
        ['SHP-20148', date(2025, 3, 18), 'Aramex',   'Mumbai',       41.5],
        ['SHP-20110', date(2025, 3, 5),  'TNT',      'Mexico City',  78.2],
        ['SHP-20096', date(2025, 2, 22), 'UPS',      'Amsterdam',    16.9],
        ['SHP-20131', date(2025, 3, 12), 'DHL',      'São Paulo',    53.3],
        ['SHP-20118', date(2025, 3, 8),  'FedEx',    'Hong Kong',    25.7],
        ['SHP-20085', date(2025, 2, 14), 'Aramex',   'Cairo',        39.0],
        ['SHP-20143', date(2025, 3, 15), 'UPS',      'Seoul',        67.8],
        ['SHP-20103', date(2025, 3, 1),  'TNT',      'Madrid',       22.4],
        ['SHP-20125', date(2025, 3, 10), 'FedEx',    'Chicago',      84.1],
        ['SHP-20090', date(2025, 2, 19), 'DHL',      'Bangkok',      11.6],
        ['SHP-20153', date(2025, 3, 20), 'Aramex',   'Lagos',        49.3],
        ['SHP-20113', date(2025, 3, 6),  'UPS',      'Rome',         36.5],
        ['SHP-20099', date(2025, 2, 25), 'FedEx',    'Johannesburg', 71.2],
        ['SHP-20138', date(2025, 3, 14), 'TNT',      'Istanbul',     28.9],
        ['SHP-20082', date(2025, 2, 11), 'DHL',      'Buenos Aires', 45.6],
        ['SHP-20146', date(2025, 3, 17), 'Aramex',   'Riyadh',       58.3],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set date number format for Dispatch Date column
    for r in range(2, len(data) + 2):
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'

    # Column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
