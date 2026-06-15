"""
Reward Script: Sort logistics shipment data by dispatch date and create line chart
Task ID: osworld_calc_sort_date_chart_008
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.40): Shipments sheet sorted by Dispatch Date ascending
  - Component 2 (0.35): Daily Counts aggregation sheet with correct date-grouped shipment counts
  - Component 3 (0.25): Line chart present showing daily shipment count over time
"""

import os
import openpyxl
from datetime import datetime

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sort_date_chart_008'


def get_chart_title_text(chart):
    """Extract plain text from chart title object."""
    try:
        if chart.title is None:
            return None
        title = chart.title
        if hasattr(title, 'tx') and title.tx is not None:
            tx = title.tx
            if hasattr(tx, 'rich') and tx.rich is not None:
                for para in tx.rich.p:
                    for r in (para.r or []):
                        return r.t
        return None
    except Exception:
        return None


def get_axis_title_text(axis):
    """Extract plain text from axis title object."""
    try:
        if axis is None or axis.title is None:
            return None
        title = axis.title
        if hasattr(title, 'tx') and title.tx is not None:
            tx = title.tx
            if hasattr(tx, 'rich') and tx.rich is not None:
                for para in tx.rich.p:
                    for r in (para.r or []):
                        return r.t
        return None
    except Exception:
        return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Shipments sheet must exist
    if 'Shipments' not in wb.sheetnames:
        print("CRITICAL: 'Shipments' sheet not found. Cannot verify.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Shipments']

    # Verify expected 25 rows (1 header + 24 data rows) with 5 columns
    if ws.max_row < 2 or ws.max_column < 2:
        print(f"CRITICAL: Shipments sheet looks empty (rows={ws.max_row}, cols={ws.max_column})")
        print("REWARD: 0.0")
        return 0.0

    # ------------------------------------------------------------------
    # Component 1: Shipments sorted by Dispatch Date ascending (0.40 pts)
    # This FAILS on initial (data is in random order) and PASSES on golden
    # Expected first date: 2025-02-11, last date: 2025-03-20
    # ------------------------------------------------------------------
    try:
        # Collect dispatch dates from column B (index 2), skipping header row
        dates = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2, values_only=True):
            val = row[0]
            if isinstance(val, datetime):
                dates.append(val)
            elif val is not None:
                # Try parsing if it's a string
                try:
                    dates.append(datetime.strptime(str(val), '%Y-%m-%d'))
                except Exception:
                    dates.append(None)

        # Filter None values
        valid_dates = [d for d in dates if d is not None]

        if len(valid_dates) < 2:
            print(f"FAIL: Component 1 — Not enough date values found (got {len(valid_dates)})")
        else:
            # Check if dates are sorted ascending
            is_sorted = all(valid_dates[i] <= valid_dates[i + 1] for i in range(len(valid_dates) - 1))
            # Also verify first and last dates match expected range
            first_date = valid_dates[0]
            last_date = valid_dates[-1]
            expected_first = datetime(2025, 2, 11)
            expected_last = datetime(2025, 3, 20)

            if is_sorted and first_date == expected_first and last_date == expected_last:
                print(f"PASS: Component 1 — Shipments sorted ascending by Dispatch Date "
                      f"({len(valid_dates)} records, {first_date.date()} to {last_date.date()}) (0.40 pts)")
                total_score += 0.40
            elif is_sorted:
                print(f"FAIL: Component 1 — Data is sorted but date range mismatch. "
                      f"Got first={first_date.date()}, last={last_date.date()}, "
                      f"expected first={expected_first.date()}, last={expected_last.date()}")
            else:
                # Show the first out-of-order pair
                for i in range(len(valid_dates) - 1):
                    if valid_dates[i] > valid_dates[i + 1]:
                        print(f"FAIL: Component 1 — Dates not sorted ascending. "
                              f"Row {i+2} ({valid_dates[i].date()}) > Row {i+3} ({valid_dates[i+1].date()})")
                        break
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Daily Counts aggregation sheet with correct values (0.35 pts)
    # This FAILS on initial (no Daily Counts sheet) and PASSES on golden
    # Expected: 24 unique dates, March 14 has count=2, all others count=1
    # ------------------------------------------------------------------
    try:
        # Find a sheet that aggregates daily shipment counts
        # It may be named 'Daily Counts' or similar
        daily_sheet = None
        for sname in wb.sheetnames:
            if sname != 'Shipments':
                candidate = wb[sname]
                # Check if it has Date and count columns
                header_row = [cell.value for cell in candidate[1]]
                if any('date' in str(h).lower() for h in header_row if h):
                    daily_sheet = candidate
                    break

        if daily_sheet is None:
            print("FAIL: Component 2 — No daily aggregation sheet found (looked for non-Shipments sheet with 'date' header)")
        else:
            # Read aggregation data
            rows = list(daily_sheet.iter_rows(min_row=2, max_row=daily_sheet.max_row, values_only=True))
            # Filter out empty rows
            data_rows = [(row[0], row[1]) for row in rows if row[0] is not None and row[1] is not None]

            # Expected: 24 unique dates
            expected_date_count = 24
            # Key spot check: March 14, 2025 should have count = 2
            march14 = datetime(2025, 3, 14)
            march14_count = None
            for date_val, count_val in data_rows:
                if isinstance(date_val, datetime) and date_val == march14:
                    march14_count = count_val
                    break

            total_count_sum = sum(c for _, c in data_rows if isinstance(c, (int, float)))
            # Total shipments = 25, so sum of counts should be 25 (24 data rows? No -- 24 unique dates, sum = 25 since Mar 14 has 2)
            # Actually: 24 records in the initial with unique dates except March 14 has 2 (SHP-20140 and SHP-20138)
            # So 24 unique dates? Let's check: 25 total records - 1 duplicate date = 24 unique dates
            # Sum of all counts = 25

            date_count_ok = len(data_rows) == expected_date_count
            total_ok = total_count_sum == 25
            march14_ok = march14_count == 2

            if date_count_ok and total_ok and march14_ok:
                print(f"PASS: Component 2 — Daily Counts aggregation correct: {len(data_rows)} unique dates, "
                      f"total={total_count_sum} shipments, March 14 count={march14_count} (0.35 pts)")
                total_score += 0.35
            else:
                issues = []
                if not date_count_ok:
                    issues.append(f"expected {expected_date_count} date rows, got {len(data_rows)}")
                if not total_ok:
                    issues.append(f"expected total count=25, got {total_count_sum}")
                if not march14_ok:
                    issues.append(f"March 14 count expected=2, got {march14_count}")
                print(f"FAIL: Component 2 — Daily Counts issues: {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Line chart present showing daily shipment count over time (0.25 pts)
    # This FAILS on initial (no charts) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        # Search all sheets for a line chart
        line_chart_found = False
        line_chart_uses_daily_data = False
        chart_has_date_axis = False

        for sname in wb.sheetnames:
            ws_check = wb[sname]
            charts = ws_check._charts
            for chart in charts:
                chart_type_name = type(chart).__name__.lower()
                if 'line' in chart_type_name:
                    line_chart_found = True
                    # Check if chart references daily count data
                    for series in chart.series:
                        if series.val is not None:
                            val_ref = str(series.val)
                            # Check that it references 'Daily Counts' or count column data
                            if 'Daily' in val_ref or 'B$2' in val_ref or 'count' in val_ref.lower():
                                line_chart_uses_daily_data = True
                    # Check x-axis title mentions date
                    if chart.x_axis is not None:
                        x_title = get_axis_title_text(chart.x_axis)
                        if x_title and 'date' in x_title.lower():
                            chart_has_date_axis = True

        if line_chart_found and line_chart_uses_daily_data:
            print(f"PASS: Component 3 — Line chart found referencing daily count data "
                  f"(date axis title present: {chart_has_date_axis}) (0.25 pts)")
            total_score += 0.25
        elif line_chart_found:
            print(f"FAIL: Component 3 — Line chart found but does not appear to reference daily count data. "
                  f"Chart found but series reference check failed.")
        else:
            print("FAIL: Component 3 — No line chart found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
