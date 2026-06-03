"""
Reward Script: Calculate remaining useful life for IT assets using DATEDIF and flag expired assets
Task ID: osworld_calc_age_calculation_datedif_008
Domain: libreoffice_calc
Scoring:
  Component 1: Column D contains DATEDIF formulas for remaining life (0.5 pts)
  Component 2: Formulas use IFERROR to handle expired assets ("0 years 0 months") (0.2 pts)
  Component 3: Conditional formatting with red background for expired assets (0.3 pts)
Total: 1.0
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_age_calculation_datedif_008'


def _get_cf_red_fill_rule(ws):
    """
    Search for a conditional formatting rule that applies a red fill.
    Returns (cf_range_str, rule, formula) or (None, None, None).
    """
    for cf_range, rules in ws.conditional_formatting._cf_rules.items():
        for rule in rules:
            if not (hasattr(rule, 'dxf') and rule.dxf is not None):
                continue
            try:
                fill = rule.dxf.fill
                if fill is None:
                    continue
                fgcolor = fill.fgColor
                if fgcolor is None:
                    continue
                rgb = fgcolor.rgb
                if rgb and rgb.upper() in ('FFFF0000', 'FF0000'):
                    formula_list = rule.formula if hasattr(rule, 'formula') and rule.formula else []
                    return (str(cf_range), rule, formula_list)
            except Exception:
                continue
    return (None, None, None)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task requirements:
    1. Column D (Remaining Life) contains DATEDIF-based formulas for all data rows
    2. Formulas handle expired assets (end date past) using IFERROR to show "0 years 0 months"
    3. Conditional formatting applies red background to rows where the asset is at or past end-of-life
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the active sheet (should be 'IT Assets')
    ws = wb.active

    # Determine data rows: row 1 is header, rows 2..max_row are data
    data_rows = list(range(2, ws.max_row + 1))
    if not data_rows:
        print("FAIL: No data rows found in spreadsheet")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Column D contains DATEDIF formulas for all data rows (0.5 pts)
    # Each formula must reference DATEDIF, TODAY(), and the expected end date
    # computed from the purchase date (B column) and expected life (C column).
    # -------------------------------------------------------------------------
    try:
        datedif_count = 0
        missing_formula_rows = []
        for r in data_rows:
            cell_d = ws.cell(row=r, column=4)
            val = cell_d.value
            if val is None:
                missing_formula_rows.append(r)
                continue
            formula_str = str(val).upper().replace(" ", "")
            if "DATEDIF" in formula_str and "TODAY()" in formula_str:
                datedif_count += 1
            else:
                missing_formula_rows.append(r)

        total_rows = len(data_rows)
        if datedif_count == total_rows:
            print(f"PASS: Component 1 — All {total_rows} data rows in column D contain DATEDIF formulas (0.5 pts)")
            total_score += 0.5
        elif datedif_count >= 1:
            partial = round(0.5 * (datedif_count / total_rows), 2)
            print(f"PARTIAL: Component 1 — {datedif_count}/{total_rows} rows contain DATEDIF formulas (+{partial} pts); missing rows: {missing_formula_rows}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {datedif_count}/{total_rows} rows have DATEDIF formulas; missing rows: {missing_formula_rows}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Formulas use IFERROR to handle expired assets (0.2 pts)
    # Formulas must wrap the DATEDIF expression in IFERROR(..., "0 years 0 months")
    # to produce a friendly message when end date is already past.
    # -------------------------------------------------------------------------
    try:
        iferror_count = 0
        for r in data_rows:
            cell_d = ws.cell(row=r, column=4)
            val = cell_d.value
            if val is None:
                continue
            formula_str = str(val).upper().replace(" ", "")
            # Check for IFERROR wrapping and "0 years 0 months" as fallback value
            # Strip quotes to compare content: "0 years 0 months" -> 0YEARS0MONTHS
            stripped = formula_str.replace('"', '').replace("'", "")
            if "IFERROR(" in formula_str and "0YEARS0MONTHS" in stripped:
                iferror_count += 1

        if iferror_count == len(data_rows):
            print(f"PASS: Component 2 — All {len(data_rows)} rows use IFERROR with '0 years 0 months' fallback (0.2 pts)")
            total_score += 0.2
        elif iferror_count >= 1:
            partial = round(0.2 * (iferror_count / len(data_rows)), 2)
            print(f"PARTIAL: Component 2 — {iferror_count}/{len(data_rows)} rows use IFERROR (+{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No rows use IFERROR with '0 years 0 months' fallback")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Conditional formatting with red background for expired assets (0.3 pts)
    # A conditional formatting rule must exist that:
    #   a) Has a red (FFFF0000) fill
    #   b) Uses a formula that checks whether the end date (purchase date + expected life) <= TODAY()
    # -------------------------------------------------------------------------
    try:
        cf_range_str, cf_rule, formula_list = _get_cf_red_fill_rule(ws)

        if cf_range_str is None:
            # No red-fill CF rule found at all
            print("FAIL: Component 3 — No conditional formatting with red background found")
            cf_debug = list(ws.conditional_formatting._cf_rules.items())
            if cf_debug:
                print(f"  Found {len(cf_debug)} CF range(s) but none with red fill")
        elif len(formula_list) == 0:
            # Red fill found but no formula
            print(f"PARTIAL: Component 3 — Red fill CF exists but no formula found (+0.15 pts)")
            total_score += 0.15
        else:
            # Check formula references end date comparison with TODAY()
            formula_checks_enddate = any(
                ("TODAY()" in f.upper().replace(" ", "") and
                 ("DATE(" in f.upper().replace(" ", "") or "$B" in f.upper()))
                for f in formula_list
            )
            if formula_checks_enddate:
                print(f"PASS: Component 3 — CF rule with red fill found, formula: {formula_list[0]} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"PARTIAL: Component 3 — Red fill CF exists but formula does not check end-of-life date; formula: {formula_list} (+0.15 pts)")
                if 0.15 > 0:
                    total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
