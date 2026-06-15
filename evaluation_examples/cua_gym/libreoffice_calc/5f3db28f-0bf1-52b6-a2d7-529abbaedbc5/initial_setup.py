"""
Initial Setup: IT Asset Remaining Life Calculation
Task ID: osworld_calc_age_calculation_datedif_008
Domain: libreoffice_calc

Creates an IT asset tracking spreadsheet with:
- Asset ID, Purchase Date, Expected Life Years columns (filled with realistic data)
- Remaining Life column (D) is empty — agent must fill with DATEDIF formulas
- No conditional formatting applied yet — agent must add red highlighting for expired assets
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_age_calculation_datedif_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: IT Assets ---
    ws = wb.active
    ws.title = "IT Assets"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Column headers
    headers = ["Asset ID", "Purchase Date", "Expected Life (Years)", "Remaining Life"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Realistic IT asset data
    # Mix of expired assets (past their end-of-life) and active assets
    # Purchase dates span 2015-2022, life years 3-7
    asset_data = [
        # Asset ID, Purchase Date (as date string stored as date), Expected Life Years
        ("IT-001", "2016-03-15", 5),   # End: 2021-03-15 => EXPIRED
        ("IT-002", "2018-07-22", 4),   # End: 2022-07-22 => EXPIRED
        ("IT-003", "2021-11-08", 5),   # End: 2026-11-08 => ACTIVE
        ("IT-004", "2017-01-30", 7),   # End: 2024-01-30 => EXPIRED
        ("IT-005", "2022-05-14", 4),   # End: 2026-05-14 => ACTIVE
        ("IT-006", "2015-09-03", 6),   # End: 2021-09-03 => EXPIRED
        ("IT-007", "2020-12-19", 5),   # End: 2025-12-19 => EXPIRED
        ("IT-008", "2023-02-28", 5),   # End: 2028-02-28 => ACTIVE
        ("IT-009", "2019-06-11", 4),   # End: 2023-06-11 => EXPIRED
        ("IT-010", "2022-08-05", 6),   # End: 2028-08-05 => ACTIVE
        ("IT-011", "2016-10-20", 5),   # End: 2021-10-20 => EXPIRED
        ("IT-012", "2021-03-17", 7),   # End: 2028-03-17 => ACTIVE
        ("IT-013", "2018-12-01", 4),   # End: 2022-12-01 => EXPIRED
        ("IT-014", "2023-09-10", 5),   # End: 2028-09-10 => ACTIVE
    ]

    from datetime import datetime

    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, (asset_id, purchase_date_str, life_years) in enumerate(asset_data, 2):
        purchase_date = datetime.strptime(purchase_date_str, "%Y-%m-%d").date()

        # Asset ID
        cell_a = ws.cell(row=row_idx, column=1, value=asset_id)
        cell_a.border = cell_border
        cell_a.alignment = Alignment(horizontal="center")

        # Purchase Date
        cell_b = ws.cell(row=row_idx, column=2, value=purchase_date)
        cell_b.number_format = "yyyy-mm-dd"
        cell_b.border = cell_border
        cell_b.alignment = Alignment(horizontal="center")

        # Expected Life Years
        cell_c = ws.cell(row=row_idx, column=3, value=life_years)
        cell_c.border = cell_border
        cell_c.alignment = Alignment(horizontal="center")

        # Column D is intentionally left empty (Remaining Life — agent must fill)
        cell_d = ws.cell(row=row_idx, column=4, value=None)
        cell_d.border = cell_border

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 24

    # Freeze header row
    ws.freeze_panes = "A2"

    # Row 1 height
    ws.row_dimensions[1].height = 25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
