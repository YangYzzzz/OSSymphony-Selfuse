"""
Reward Script: Team Meeting Minutes and Action Item Tracker
Task ID: calc_grs_093
Domain: libreoffice_calc
Scoring:
  Component 1: Conditional formatting on Action Items (3 rules: overdue=red, due-this-week=yellow, complete=green) — 0.30
  Component 2: Action Items sorted by Due Date ascending — 0.20
  Component 3: Auto filter applied on Action Items sheet — 0.10
  Component 4: Dashboard populated with open action items by owner — 0.20
  Component 5: Dashboard summary metrics (completion rate, overdue count) — 0.10
  Component 6: Dashboard has at least one chart — 0.10
"""

import os
import openpyxl
from datetime import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_093'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Required sheets exist
    required_sheets = ['Meeting Log', 'Action Items', 'Dashboard']
    for sheet_name in required_sheets:
        if sheet_name not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet_name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    ws_ai = wb['Action Items']
    ws_dash = wb['Dashboard']

    # ========================================================================
    # Component 1: Conditional formatting on Action Items (0.30 points)
    # Task requires: overdue=red, due this week=yellow, complete=green
    # Initial has 0 CF rules; golden has 3.
    # ========================================================================
    try:
        cf_rules_list = list(ws_ai.conditional_formatting)
        # Collect all rules across all CF ranges
        all_rules = []
        for cf in cf_rules_list:
            for rule in cf.rules:
                all_rules.append(rule)

        cf_score = 0.0
        has_red_overdue = False
        has_yellow_due_soon = False
        has_green_complete = False

        for rule in all_rules:
            formula_str = str(getattr(rule, 'formula', '')).upper()
            # Check fill color
            fill_rgb = None
            try:
                if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                    fill_rgb = rule.dxf.fill.fgColor.rgb
            except Exception:
                pass

            # Overdue rule: formula references TODAY() with < comparison, red fill
            if fill_rgb and 'FF' in fill_rgb[:2]:
                rgb_no_alpha = fill_rgb[2:] if len(fill_rgb) == 8 else fill_rgb
                # Red-ish colors (high R, low G, low B)
                try:
                    r_val = int(rgb_no_alpha[0:2], 16)
                    g_val = int(rgb_no_alpha[2:4], 16)
                    b_val = int(rgb_no_alpha[4:6], 16)
                except (ValueError, IndexError):
                    r_val, g_val, b_val = 0, 0, 0

                if r_val > 180 and g_val < 100 and b_val < 100:
                    # Red fill - likely overdue rule
                    if 'TODAY' in formula_str and 'COMPLETE' in formula_str:
                        has_red_overdue = True
                elif g_val > 100 and r_val < 100 and b_val < 120:
                    # Green fill - likely complete rule (e.g., 00B050 = R0 G176 B80)
                    if 'COMPLETE' in formula_str:
                        has_green_complete = True
                elif r_val > 180 and g_val > 180 and b_val < 100:
                    # Yellow fill - likely due this week rule
                    if 'TODAY' in formula_str:
                        has_yellow_due_soon = True

        if has_red_overdue:
            cf_score += 0.10
            print("PASS: Component 1a — Overdue items conditional formatting (red) found (0.10 pts)")
        else:
            print("FAIL: Component 1a — No red conditional formatting for overdue items")

        if has_yellow_due_soon:
            cf_score += 0.10
            print("PASS: Component 1b — Due-this-week conditional formatting (yellow) found (0.10 pts)")
        else:
            print("FAIL: Component 1b — No yellow conditional formatting for due-this-week items")

        if has_green_complete:
            cf_score += 0.10
            print("PASS: Component 1c — Complete items conditional formatting (green) found (0.10 pts)")
        else:
            print("FAIL: Component 1c — No green conditional formatting for complete items")

        total_score += cf_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ========================================================================
    # Component 2: Action Items sorted by Due Date ascending (0.20 points)
    # Initial has items NOT sorted by due date; golden has them sorted.
    # ========================================================================
    try:
        due_dates = []
        for r in range(2, ws_ai.max_row + 1):
            due_val = ws_ai.cell(r, 5).value  # Column E = Due Date
            if due_val is not None:
                if isinstance(due_val, datetime):
                    due_dates.append(due_val)
                elif isinstance(due_val, str):
                    try:
                        due_dates.append(datetime.strptime(due_val, '%Y-%m-%d'))
                    except ValueError:
                        due_dates.append(datetime.strptime(due_val, '%m/%d/%Y'))

        if len(due_dates) >= 2:
            is_sorted = all(due_dates[i] <= due_dates[i + 1] for i in range(len(due_dates) - 1))
            if is_sorted:
                print(f"PASS: Component 2 — Action Items sorted by Due Date ascending ({len(due_dates)} items) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — Action Items NOT sorted by Due Date ascending")
                # Show first out-of-order pair
                for i in range(len(due_dates) - 1):
                    if due_dates[i] > due_dates[i + 1]:
                        print(f"  First violation: row {i+2} ({due_dates[i].date()}) > row {i+3} ({due_dates[i+1].date()})")
                        break
        else:
            print(f"FAIL: Component 2 — Not enough due dates found ({len(due_dates)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ========================================================================
    # Component 3: Auto filter applied on Action Items sheet (0.10 points)
    # Initial has no auto filter; golden has A1:H16.
    # ========================================================================
    try:
        af_ref = ws_ai.auto_filter.ref
        if af_ref:
            print(f"PASS: Component 3 — Auto filter set on Action Items: {af_ref} (0.10 pts)")
            total_score += 0.10
        else:
            print("FAIL: Component 3 — No auto filter on Action Items sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ========================================================================
    # Component 4: Dashboard populated with open action items by owner (0.20 points)
    # Initial Dashboard has only a title; golden has an owner-based breakdown table.
    # Check: Dashboard has multiple rows with owner names and COUNTIFS formulas.
    # ========================================================================
    try:
        # Look for owner names in Dashboard column A (rows 2+)
        owner_names_found = 0
        countifs_found = 0
        for r in range(2, ws_dash.max_row + 1):
            a_val = ws_dash.cell(r, 1).value
            b_val = ws_dash.cell(r, 2).value
            if a_val and isinstance(a_val, str) and len(a_val.strip()) > 0:
                # Check if this row has a person name (not a section header)
                b_str = str(b_val) if b_val else ''
                if 'COUNTIF' in b_str.upper():
                    owner_names_found += 1
                    countifs_found += 1

        dash_score = 0.0
        if owner_names_found >= 4:
            # At least 4 unique owners with COUNTIFS formulas
            dash_score = 0.20
            print(f"PASS: Component 4 — Dashboard has {owner_names_found} owners with COUNTIFS formulas (0.20 pts)")
        elif owner_names_found >= 2:
            dash_score = 0.10
            print(f"PARTIAL: Component 4 — Dashboard has {owner_names_found} owners (expected >= 4) (0.10 pts)")
        else:
            print(f"FAIL: Component 4 — Dashboard has {owner_names_found} owners with COUNTIFS formulas (expected >= 4)")

        total_score += dash_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ========================================================================
    # Component 5: Dashboard summary metrics (0.10 points)
    # Golden has: Total Action Items, Completed Items, Completion Rate, Overdue Items
    # Initial Dashboard has none of these.
    # ========================================================================
    try:
        # Search for key metric labels and formulas in Dashboard
        has_completion_rate = False
        has_overdue_count = False

        for r in range(2, ws_dash.max_row + 1):
            a_val = str(ws_dash.cell(r, 1).value or '').lower()
            b_val = str(ws_dash.cell(r, 2).value or '')

            if 'completion' in a_val and 'rate' in a_val:
                # Check B has a formula
                if '=' in b_val or (ws_dash.cell(r, 2).value is not None and isinstance(ws_dash.cell(r, 2).value, (int, float))):
                    has_completion_rate = True

            if 'overdue' in a_val:
                if '=' in b_val or (ws_dash.cell(r, 2).value is not None and isinstance(ws_dash.cell(r, 2).value, (int, float))):
                    has_overdue_count = True

        metrics_score = 0.0
        if has_completion_rate and has_overdue_count:
            metrics_score = 0.10
            print("PASS: Component 5 — Dashboard has completion rate and overdue count metrics (0.10 pts)")
        elif has_completion_rate or has_overdue_count:
            metrics_score = 0.05
            found = 'completion rate' if has_completion_rate else 'overdue count'
            missing = 'overdue count' if has_completion_rate else 'completion rate'
            print(f"PARTIAL: Component 5 — Dashboard has {found} but missing {missing} (0.05 pts)")
        else:
            print("FAIL: Component 5 — Dashboard missing completion rate and overdue count metrics")

        total_score += metrics_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ========================================================================
    # Component 6: Dashboard has at least one chart (0.10 points)
    # Initial Dashboard has 0 charts; golden has 1.
    # ========================================================================
    try:
        chart_count = len(ws_dash._charts)
        if chart_count >= 1:
            print(f"PASS: Component 6 — Dashboard has {chart_count} chart(s) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — Dashboard has no charts (expected >= 1)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
