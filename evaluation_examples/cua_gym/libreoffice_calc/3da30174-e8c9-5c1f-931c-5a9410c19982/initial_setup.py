"""
Initial Setup: Team meeting minutes and action item tracker
Task ID: calc_grs_093
Domain: libreoffice_calc

Creates a workbook with Meeting Log, Action Items, and Dashboard sheets.
The initial state has raw data with dropdowns but NO conditional formatting,
NO dashboard formulas/charts, NO sorting, and NO auto-filter.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # =========================================================
    # Sheet 1: Meeting Log
    # =========================================================
    ws_meet = wb.active
    ws_meet.title = "Meeting Log"

    meet_headers = ["Meeting Date", "Meeting Type", "Attendees", "Facilitator",
                    "Key Decisions"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(meet_headers, 1):
        cell = ws_meet.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Meeting data (8 meetings over ~2 months)
    meetings = [
        ["2026-02-03", "Team Standup", "All team (8)", "Sarah Chen",
         "Agreed to adopt new sprint planning format; Will migrate Jira boards by Friday"],
        ["2026-02-10", "Department Review", "Sarah Chen, Marcus Johnson, Priya Patel, James Wilson, Emily Zhang",
         "Marcus Johnson",
         "Q4 results exceeded target by 12%; Budget reallocation approved for cloud migration"],
        ["2026-02-17", "Project Update", "Sarah Chen, Priya Patel, Tomoko Yamada, Luis Garcia",
         "Priya Patel",
         "Phase 2 delivery moved to March 15; Need additional QA resource"],
        ["2026-02-24", "Team Standup", "All team (8)", "James Wilson",
         "Holiday schedule confirmed; Cover assignments for March PTO"],
        ["2026-03-03", "Strategy", "Sarah Chen, Marcus Johnson, Emily Zhang, David Kim",
         "Sarah Chen",
         "New client onboarding process approved; Target 20% reduction in onboarding time"],
        ["2026-03-10", "1on1", "Sarah Chen, Marcus Johnson", "Sarah Chen",
         "Marcus to lead Q1 retrospective; Performance review prep by March 20"],
        ["2026-03-17", "Project Update", "All team (8)", "Tomoko Yamada",
         "Phase 2 testing complete; Go-live scheduled for March 28"],
        ["2026-03-24", "Team Standup", "All team (8)", "Luis Garcia",
         "Sprint velocity improved 15%; Backlog grooming session scheduled for April 1"],
    ]

    for r, row_data in enumerate(meetings, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_meet.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    ws_meet.column_dimensions["A"].width = 15
    ws_meet.column_dimensions["B"].width = 20
    ws_meet.column_dimensions["C"].width = 40
    ws_meet.column_dimensions["D"].width = 18
    ws_meet.column_dimensions["E"].width = 50

    # Data validation for Meeting Type dropdown
    dv_type = DataValidation(
        type="list",
        formula1='"Team Standup,Department Review,Project Update,Strategy,1on1"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_type.error = "Please select a valid meeting type"
    dv_type.errorTitle = "Invalid Meeting Type"
    dv_type.add("B2:B100")
    ws_meet.add_data_validation(dv_type)

    # =========================================================
    # Sheet 2: Action Items (NOT sorted, NO conditional formatting, NO filter)
    # =========================================================
    ws_actions = wb.create_sheet("Action Items")

    action_headers = ["Action ID", "Description", "Owner", "Priority",
                      "Due Date", "Status", "Linked Meeting Date", "Notes"]

    for col, h in enumerate(action_headers, 1):
        cell = ws_actions.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Action items data (15 items, deliberately NOT sorted by due date)
    actions = [
        ["AI-001", "Migrate Jira boards to new sprint format", "James Wilson", "High",
         "2026-02-07", "Complete", "2026-02-03", "Completed on Feb 6"],
        ["AI-002", "Prepare Q4 results presentation for leadership", "Marcus Johnson", "High",
         "2026-02-14", "Complete", "2026-02-10", "Presented to VP on Feb 13"],
        ["AI-003", "Hire additional QA resource for Phase 2", "Sarah Chen", "High",
         "2026-03-01", "In Progress", "2026-02-17", "Interviews scheduled for week of Feb 24"],
        ["AI-004", "Create holiday cover assignment matrix", "Emily Zhang", "Medium",
         "2026-02-28", "Complete", "2026-02-24", "Shared via email Feb 27"],
        ["AI-005", "Draft new client onboarding playbook", "David Kim", "High",
         "2026-03-14", "In Progress", "2026-03-03", "First draft under review"],
        ["AI-006", "Benchmark current onboarding time metrics", "Priya Patel", "Medium",
         "2026-03-10", "Complete", "2026-03-03", "Baseline: avg 14 days"],
        ["AI-007", "Prepare Q1 retrospective agenda and materials", "Marcus Johnson", "Medium",
         "2026-03-20", "Not Started", "2026-03-10", ""],
        ["AI-008", "Complete performance review self-assessment", "Marcus Johnson", "High",
         "2026-03-20", "Not Started", "2026-03-10", ""],
        ["AI-009", "Run Phase 2 go-live readiness checklist", "Tomoko Yamada", "High",
         "2026-03-26", "In Progress", "2026-03-17", "80% items checked off"],
        ["AI-010", "Coordinate go-live support rotation schedule", "Luis Garcia", "Medium",
         "2026-03-25", "Not Started", "2026-03-17", ""],
        ["AI-011", "Schedule backlog grooming session with product team", "Luis Garcia", "Low",
         "2026-03-28", "Not Started", "2026-03-24", ""],
        ["AI-012", "Update sprint velocity dashboard with Q1 data", "James Wilson", "Medium",
         "2026-03-31", "Not Started", "2026-03-24", ""],
        ["AI-013", "Review and update cloud migration budget estimates", "Sarah Chen", "High",
         "2026-02-21", "Complete", "2026-02-10", "Updated estimates submitted Feb 20"],
        ["AI-014", "Document Phase 2 test results and lessons learned", "Tomoko Yamada", "Medium",
         "2026-03-21", "In Progress", "2026-03-17", "Draft 60% complete"],
        ["AI-015", "Set up automated regression test suite for Phase 2", "Priya Patel", "High",
         "2026-03-24", "Cancelled", "2026-02-17",
         "Descoped - using manual testing for go-live; will automate in Phase 3"],
    ]

    for r, row_data in enumerate(actions, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_actions.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Column widths
    ws_actions.column_dimensions["A"].width = 12
    ws_actions.column_dimensions["B"].width = 45
    ws_actions.column_dimensions["C"].width = 18
    ws_actions.column_dimensions["D"].width = 12
    ws_actions.column_dimensions["E"].width = 14
    ws_actions.column_dimensions["F"].width = 14
    ws_actions.column_dimensions["G"].width = 18
    ws_actions.column_dimensions["H"].width = 40

    # Data validation for Priority dropdown
    dv_priority = DataValidation(
        type="list",
        formula1='"High,Medium,Low"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_priority.add("D2:D100")
    ws_actions.add_data_validation(dv_priority)

    # Data validation for Status dropdown
    dv_status = DataValidation(
        type="list",
        formula1='"Not Started,In Progress,Complete,Cancelled"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_status.add("F2:F100")
    ws_actions.add_data_validation(dv_status)

    # =========================================================
    # Sheet 3: Dashboard (empty - task asks user to create it)
    # =========================================================
    ws_dash = wb.create_sheet("Dashboard")

    # Just a title placeholder - no formulas, no charts
    cell = ws_dash.cell(row=1, column=1, value="Team Meeting Action Items Dashboard")
    cell.font = Font(name="Calibri", size=14, bold=True)
    ws_dash.column_dimensions["A"].width = 40

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
