"""
Reward Script: Re-import CSV with UTF-8 encoding to fix garbled accented characters
Task ID: calc_tbl_031
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Names in column B contain correct accented characters (no mojibake)
  Component 2 (0.3): Locations in column G contain correct accented characters (no mojibake)
  Component 3 (0.3): Data integrity - all 12 employee rows present with correct structure
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_031'

# Known correct values for names (column B) - these have accented chars
EXPECTED_NAMES = {
    2: 'Hans Müller',
    3: 'Renée Dupont',
    4: 'José García',
    5: 'Søren Andersen',
    6: 'François Lefèvre',
    7: 'Björk Jónsdóttir',
    8: 'Zoë Mühlenberg',
    9: 'Adrián Peña',
    10: 'Hélène Beaumont',
    11: 'Günther Weiß',
    12: 'Núria Aragonés',
    13: 'Ólafur Sigurðsson',
}

# Known correct values for locations (column G) - subset with accented chars
EXPECTED_LOCATIONS = {
    2: 'München',
    5: 'København',
    7: 'Reykjavík',
    8: 'Zürich',
    9: 'São Paulo',
    10: 'Montréal',
    13: 'Reykjavík',
}

# Mojibake signatures - if any of these byte sequences appear, encoding is wrong
MOJIBAKE_MARKERS = ['Ã¼', 'Ã©', 'Ã³', 'Ã¸', 'Ã§', 'Ã¶', 'Ã«', 'Ã±', 'Ã¡', 'Ã', 'Ã°', 'Ã\x9f']


def has_mojibake(text):
    """Check if text contains mojibake markers (garbled UTF-8-as-Latin1)."""
    if text is None:
        return False
    s = str(text)
    for marker in MOJIBAKE_MARKERS:
        if marker in s:
            return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the sheet with employee data
    ws = None
    for name in wb.sheetnames:
        sheet = wb[name]
        if sheet.cell(row=1, column=1).value == 'Employee ID':
            ws = sheet
            break

    if ws is None:
        # Try the active sheet
        ws = wb.active
        if ws.cell(row=1, column=1).value != 'Employee ID':
            print("FAIL: No sheet found with 'Employee ID' header in A1")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Names in column B have correct accented characters (0.4 points)
    # This checks that mojibake is gone AND correct characters are present
    try:
        names_correct = 0
        names_total = len(EXPECTED_NAMES)

        for row, expected_name in EXPECTED_NAMES.items():
            actual = ws.cell(row=row, column=2).value
            if actual is None:
                print(f"  FAIL: B{row} is empty, expected '{expected_name}'")
                continue

            actual_str = str(actual).strip()
            # Check: no mojibake AND matches expected
            if not has_mojibake(actual_str) and actual_str == expected_name:
                names_correct += 1
            else:
                print(f"  FAIL: B{row} = '{actual_str}', expected '{expected_name}'")

        if names_correct == names_total:
            print(f"PASS: Component 1 — All {names_total} names have correct accented characters (0.4 pts)")
            total_score += 0.4
        elif names_correct > 0:
            partial = round(0.4 * (names_correct / names_total), 2)
            print(f"PARTIAL: Component 1 — {names_correct}/{names_total} names correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — 0/{names_total} names have correct encoding")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Locations in column G have correct accented characters (0.3 points)
    # Only checks locations that contain accented chars (the ones that would be garbled)
    try:
        locs_correct = 0
        locs_total = len(EXPECTED_LOCATIONS)

        for row, expected_loc in EXPECTED_LOCATIONS.items():
            actual = ws.cell(row=row, column=7).value
            if actual is None:
                print(f"  FAIL: G{row} is empty, expected '{expected_loc}'")
                continue

            actual_str = str(actual).strip()
            if not has_mojibake(actual_str) and actual_str == expected_loc:
                locs_correct += 1
            else:
                print(f"  FAIL: G{row} = '{actual_str}', expected '{expected_loc}'")

        if locs_correct == locs_total:
            print(f"PASS: Component 2 — All {locs_total} accented locations correct (0.3 pts)")
            total_score += 0.3
        elif locs_correct > 0:
            partial = round(0.3 * (locs_correct / locs_total), 2)
            print(f"PARTIAL: Component 2 — {locs_correct}/{locs_total} locations correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — 0/{locs_total} accented locations correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data integrity - all 12 rows present with correct non-accented data (0.3 points)
    # Verify row count AND that non-name/location columns are intact
    # This component only passes when accented data is also correct (anchored to task change)
    try:
        # Count data rows (excluding header)
        data_rows = 0
        for r in range(2, 100):
            if ws.cell(row=r, column=1).value is not None:
                data_rows += 1
            else:
                break

        # Check structural integrity: 12 rows, 7 columns, AND no mojibake anywhere
        mojibake_count = sum(
            1 for r in range(2, min(data_rows + 2, 14))
            for c in range(1, 8)
            if ws.cell(row=r, column=c).value is not None
            and has_mojibake(str(ws.cell(row=r, column=c).value))
        )

        if data_rows == 12 and mojibake_count == 0:
            print(f"PASS: Component 3 — 12 data rows present with no mojibake anywhere (0.3 pts)")
            total_score += 0.3
        elif data_rows == 12 and mojibake_count > 0:
            print(f"FAIL: Component 3 — 12 rows present but {mojibake_count} cells still have mojibake")
        else:
            print(f"FAIL: Component 3 — Expected 12 data rows, found {data_rows}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Main execution
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
