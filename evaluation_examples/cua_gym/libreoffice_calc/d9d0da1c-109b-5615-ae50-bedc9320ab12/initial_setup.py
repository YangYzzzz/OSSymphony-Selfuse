"""
Initial Setup: CSV imported with wrong encoding showing garbled accented names
Task ID: calc_tbl_031
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_031'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
CSV_DIR = '/root/data'
CSV_PATH = f'{CSV_DIR}/employees.csv'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_csv():
    """Create the UTF-8 encoded CSV source file at /root/data/employees.csv."""
    csv_content = """Employee ID,Full Name,Department,Position,Salary,Start Date,Location
E001,Hans Müller,Engineering,Senior Developer,92000,2021-03-15,München
E002,Renée Dupont,Marketing,Campaign Manager,78500,2022-01-10,Paris
E003,José García,Sales,Regional Director,95000,2020-06-22,Madrid
E004,Søren Andersen,Finance,Senior Analyst,84000,2021-09-01,København
E005,François Lefèvre,Engineering,Tech Lead,98500,2019-11-15,Lyon
E006,Björk Jónsdóttir,HR,Talent Acquisition,72000,2023-02-28,Reykjavík
E007,Zoë Mühlenberg,Operations,Logistics Manager,81000,2022-07-14,Zürich
E008,Adrián Peña,Sales,Account Executive,69500,2023-05-20,São Paulo
E009,Hélène Beaumont,Marketing,Brand Strategist,76000,2021-12-01,Montréal
E010,Günther Weiß,Finance,Controller,105000,2018-04-10,Wien
E011,Núria Aragonés,Engineering,QA Lead,82000,2022-03-18,Barcelona
E012,Ólafur Sigurðsson,Operations,Supply Chain,74500,2021-08-25,Reykjavík
"""
    # Write CSV to a temp location first, then use sudo to place it
    tmp_csv = '/tmp/employees.csv'
    with open(tmp_csv, 'w', encoding='utf-8') as f:
        f.write(csv_content)
    # Use sudo to create the target directory and move the file
    subprocess.run('echo "password" | sudo -S mkdir -p ' + CSV_DIR, shell=True, check=True,
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    subprocess.run('echo "password" | sudo -S cp ' + tmp_csv + ' ' + CSV_PATH, shell=True, check=True,
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    subprocess.run('echo "password" | sudo -S chmod 644 ' + CSV_PATH, shell=True, check=True,
                   stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    print(f'CSV file created: {CSV_PATH}')


def create_garbled_xlsx():
    """
    Create an xlsx that simulates wrong-encoding CSV import.
    We read the UTF-8 CSV bytes as Latin-1, producing mojibake.
    """
    # Read the CSV from the temp copy (user can access /tmp)
    with open('/tmp/employees.csv', 'rb') as f:
        raw = f.read()
    garbled_text = raw.decode('latin-1')
    lines = garbled_text.strip().split('\n')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "employees"

    # Header styling
    header_font = Font(name="Liberation Sans", size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for r_idx, line in enumerate(lines):
        cols = line.split(',')
        for c_idx, val in enumerate(cols):
            cell = ws.cell(row=r_idx + 1, column=c_idx + 1, value=val.strip())
            cell.border = thin_border
            if r_idx == 0:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            else:
                cell.font = Font(name="Liberation Sans", size=11)
                # Make salary numeric where possible
                if c_idx == 4:  # Salary column
                    try:
                        cell.value = int(val.strip())
                        cell.number_format = '#,##0'
                    except ValueError:
                        pass

    # Adjust column widths for readability
    col_widths = {'A': 14, 'B': 26, 'C': 16, 'D': 22, 'E': 12, 'F': 14, 'G': 18}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Garbled xlsx created: {OUTPUT}')


def main():
    create_csv()
    create_garbled_xlsx()
    # Open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


main()
