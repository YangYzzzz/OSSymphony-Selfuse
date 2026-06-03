"""
Initial Setup: Panel Speakers spreadsheet for systems conference
Task ID: osworld_multi_apps_web_prof_email_011
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Creates Panel_Speakers.xlsx with 8 panelists.
Columns: Full Name, Homepage, Email, Top Award, Institution
Email, Top Award, and Institution columns are intentionally left blank.
Chrome is also launched for web browsing.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_011'
OUTPUT = f'{WORKDIR}/Panel_Speakers.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Panel_Speakers"

    # Column headers
    headers = ["Full Name", "Homepage", "Email", "Top Award", "Institution"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, name="Calibri", size=11)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, name="Calibri", size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 8 panelists - believable systems/CS researchers
    # Unsorted initial order (not yet alphabetical by last name)
    # Email, Top Award, Institution are intentionally BLANK
    panelists = [
        # Full Name, Homepage, Email (blank), Top Award (blank), Institution (blank)
        ["Matei Zaharia",    "https://people.eecs.berkeley.edu/~matei/",        "", "", ""],
        ["Fei-Fei Li",       "https://profiles.stanford.edu/fei-fei-li",        "", "", ""],
        ["David Patterson",  "https://www2.eecs.berkeley.edu/Faculty/Homepages/patterson.html", "", "", ""],
        ["Srinivasan Keshav","https://svr-sk818-web.cl.cam.ac.uk/keshav/wiki/index.php/Main_Page", "", "", ""],
        ["Jennifer Rexford", "https://www.cs.princeton.edu/~jrex/",             "", "", ""],
        ["Ion Stoica",       "https://people.eecs.berkeley.edu/~istoica/",      "", "", ""],
        ["Hari Balakrishnan","http://nms.csail.mit.edu/~hari/",                 "", "", ""],
        ["Arvind Krishnamurthy", "https://www.cs.washington.edu/people/faculty/arvind", "", "", ""],
    ]

    for row_idx, panelist in enumerate(panelists, 2):
        for col_idx, value in enumerate(panelist, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Calibri", size=11)
            cell.alignment = Alignment(vertical="center")

    # Set column widths
    ws.column_dimensions["A"].width = 28  # Full Name
    ws.column_dimensions["B"].width = 55  # Homepage
    ws.column_dimensions["C"].width = 35  # Email
    ws.column_dimensions["D"].width = 40  # Top Award
    ws.column_dimensions["E"].width = 35  # Institution

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup: open Chrome first, then LibreOffice Calc
    launch_gui('google-chrome --new-window "about:blank"', delay_sec=2.0)
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0")


create_initial()
