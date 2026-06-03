"""
Reward Script: Panel Speakers Spreadsheet - Fill Email/Award/Institution and Sort
Task ID: osworld_multi_apps_web_prof_email_011
Domain: libreoffice_calc
Scoring:
  Component 1: Email column filled for all 8 panelists      (0.30 pts)
  Component 2: Top Award column filled for all 8 panelists  (0.25 pts)
  Component 3: Institution column filled for all 8 panelists (0.25 pts)
  Component 4: Rows sorted alphabetically by last name       (0.20 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_011'
FILE_PATH = f'{WORKDIR}/Panel_Speakers.xlsx'

# Golden reference values for all 8 panelists (keyed by full name)
# These are the expected values in the golden environment.
GOLDEN_DATA = {
    'Hari Balakrishnan':    ('hari@csail.mit.edu',         'ACM Fellow',                    'Massachusetts Institute of Technology'),
    'Srinivasan Keshav':    ('keshav@netwinder.org',        'ACM Fellow',                    'University of Cambridge'),
    'Arvind Krishnamurthy': ('arvind@cs.washington.edu',    'ACM Fellow',                    'University of Washington'),
    'Fei-Fei Li':           ('feifeili@cs.stanford.edu',    'MacArthur Fellow',              'Stanford University'),
    'David Patterson':      ('pattrsn@eecs.berkeley.edu',   'ACM A.M. Turing Award',         'University of California, Berkeley'),
    'Jennifer Rexford':     ('jrex@cs.princeton.edu',       'ACM Fellow',                    'Princeton University'),
    'Ion Stoica':           ('istoica@cs.berkeley.edu',     'ACM Fellow',                    'University of California, Berkeley'),
    'Matei Zaharia':        ('matei@cs.stanford.edu',       'ACM Grace Murray Hopper Award', 'Stanford University'),
}

# Expected sort order by last name (alphabetical)
EXPECTED_ORDER = [
    'Hari Balakrishnan',
    'Srinivasan Keshav',
    'Arvind Krishnamurthy',
    'Fei-Fei Li',
    'David Patterson',
    'Jennifer Rexford',
    'Ion Stoica',
    'Matei Zaharia',
]


def get_last_name(full_name):
    """Extract last name from full name (last space-separated token)."""
    if not full_name:
        return ''
    parts = str(full_name).strip().split()
    return parts[-1] if parts else ''


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the active sheet
    ws = wb.active

    # Verify expected headers (precondition gate — no score)
    try:
        headers = [ws.cell(1, c).value for c in range(1, 6)]
        expected_headers = ['Full Name', 'Homepage', 'Email', 'Top Award', 'Institution']
        if headers != expected_headers:
            print(f"CRITICAL: Unexpected headers: {headers}. Expected {expected_headers}")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"CRITICAL: Cannot read headers: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Read all data rows (rows 2..9)
    rows = {}
    try:
        for row_idx in range(2, ws.max_row + 1):
            full_name = ws.cell(row_idx, 1).value
            email = ws.cell(row_idx, 3).value
            award = ws.cell(row_idx, 4).value
            institution = ws.cell(row_idx, 5).value
            if full_name:
                rows[str(full_name).strip()] = {
                    'row': row_idx,
                    'email': email,
                    'award': award,
                    'institution': institution,
                }
        print(f"INFO: Found {len(rows)} data rows")
    except Exception as e:
        print(f"CRITICAL: Cannot read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify we have exactly 8 data rows (precondition gate)
    if len(rows) != 8:
        print(f"CRITICAL: Expected 8 data rows, found {len(rows)}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Email column filled for all 8 panelists (0.30 points)
    # In the initial env, all emails are None. In golden, all are filled.
    # -----------------------------------------------------------------------
    try:
        email_filled_count = 0
        email_correct_count = 0
        for name, data in rows.items():
            email = data['email']
            if email is not None and str(email).strip() != '':
                email_filled_count += 1
                # Also check against expected golden values
                if name in GOLDEN_DATA:
                    expected_email = GOLDEN_DATA[name][0]
                    if str(email).strip().lower() == expected_email.lower():
                        email_correct_count += 1

        if email_filled_count == 8:
            if email_correct_count >= 6:
                # All filled and most match expected values
                print(f"PASS: Component 1 — Email filled for all 8 rows, {email_correct_count}/8 match expected values (0.30 pts)")
                total_score += 0.30
            elif email_correct_count >= 4:
                # All filled but fewer correct
                print(f"PARTIAL: Component 1 — Email filled for all 8 rows, only {email_correct_count}/8 match expected values (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 1 — Email filled for all 8 rows but only {email_correct_count}/8 match expected values")
        else:
            print(f"FAIL: Component 1 — Email filled for only {email_filled_count}/8 rows")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Top Award column filled for all 8 panelists (0.25 points)
    # In the initial env, all awards are None. In golden, all are filled.
    # -----------------------------------------------------------------------
    try:
        award_filled_count = 0
        award_correct_count = 0
        for name, data in rows.items():
            award = data['award']
            if award is not None and str(award).strip() != '':
                award_filled_count += 1
                if name in GOLDEN_DATA:
                    expected_award = GOLDEN_DATA[name][1]
                    if str(award).strip().lower() == expected_award.lower():
                        award_correct_count += 1

        if award_filled_count == 8:
            if award_correct_count >= 6:
                print(f"PASS: Component 2 — Top Award filled for all 8 rows, {award_correct_count}/8 match expected values (0.25 pts)")
                total_score += 0.25
            elif award_correct_count >= 4:
                print(f"PARTIAL: Component 2 — Top Award filled for all 8 rows, only {award_correct_count}/8 match expected values (0.12 pts)")
                total_score += 0.12
            else:
                print(f"FAIL: Component 2 — Top Award filled for all 8 rows but only {award_correct_count}/8 match expected values")
        else:
            print(f"FAIL: Component 2 — Top Award filled for only {award_filled_count}/8 rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Institution column filled for all 8 panelists (0.25 points)
    # In the initial env, all institutions are None. In golden, all are filled.
    # -----------------------------------------------------------------------
    try:
        inst_filled_count = 0
        inst_correct_count = 0
        for name, data in rows.items():
            institution = data['institution']
            if institution is not None and str(institution).strip() != '':
                inst_filled_count += 1
                if name in GOLDEN_DATA:
                    expected_inst = GOLDEN_DATA[name][2]
                    if str(institution).strip().lower() == expected_inst.lower():
                        inst_correct_count += 1

        if inst_filled_count == 8:
            if inst_correct_count >= 6:
                print(f"PASS: Component 3 — Institution filled for all 8 rows, {inst_correct_count}/8 match expected values (0.25 pts)")
                total_score += 0.25
            elif inst_correct_count >= 4:
                print(f"PARTIAL: Component 3 — Institution filled for all 8 rows, only {inst_correct_count}/8 match expected values (0.12 pts)")
                total_score += 0.12
            else:
                print(f"FAIL: Component 3 — Institution filled for all 8 rows but only {inst_correct_count}/8 match expected values")
        else:
            print(f"FAIL: Component 3 — Institution filled for only {inst_filled_count}/8 rows")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Rows sorted alphabetically by last name (0.20 points)
    # In the initial env, rows are in the original unsorted order.
    # In the golden env, rows are sorted alphabetically by last name.
    # -----------------------------------------------------------------------
    try:
        # Collect full names in their row order (rows 2..9)
        ordered_names = []
        for row_idx in range(2, ws.max_row + 1):
            name_val = ws.cell(row_idx, 1).value
            if name_val:
                ordered_names.append(str(name_val).strip())

        # Extract last names in current order
        last_names_actual = [get_last_name(n) for n in ordered_names]

        # Expected sorted last names
        last_names_expected = [get_last_name(n) for n in EXPECTED_ORDER]

        if last_names_actual == last_names_expected:
            print(f"PASS: Component 4 — Rows sorted alphabetically by last name: {last_names_actual} (0.20 pts)")
            total_score += 0.20
        else:
            # Check if at least the order is a valid alphabetical sort
            sorted_last_names = sorted(last_names_actual)
            if last_names_actual == sorted_last_names:
                print(f"PASS: Component 4 — Rows are in valid alphabetical order by last name (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Rows NOT sorted by last name. Actual: {last_names_actual}, Expected: {last_names_expected}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
