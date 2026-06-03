"""
Initial Setup: Employee onboarding form with no validations
Task ID: calc_nrv_076
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Onboarding ---
    ws = wb.active
    ws.title = 'Onboarding'

    # Title row
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Employee Onboarding Form'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Headers
    ws['A2'] = 'Field'
    ws['B2'] = 'Value'
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    for cell_ref in ['A2', 'B2']:
        ws[cell_ref].font = header_font
        ws[cell_ref].fill = header_fill
        ws[cell_ref].alignment = Alignment(horizontal='center')

    # Form fields
    form_data = [
        ('Hire Date', None),
        ('Department', None),
        ('Starting Salary', None),
        ('Employee Name', None),
        ('Employee ID', None),
        ('Manager', None),
        ('Office Location', None),
        ('Employment Type', None),
    ]

    label_font = Font(name='Arial', size=11, bold=True)
    thin_side = Side(style='thin', color='000000')
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for i, (label, value) in enumerate(form_data, 3):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=1).border = cell_border
        ws.cell(row=i, column=2, value=value)
        ws.cell(row=i, column=2).border = cell_border

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 30

    # Row heights
    ws.row_dimensions[1].height = 35

    # --- Sheet 2: Config ---
    ws2 = wb.create_sheet('Config')
    ws2['A1'] = 'Engineering'
    ws2['A2'] = 'Marketing'
    ws2['A3'] = 'Finance'
    ws2['A4'] = 'Human Resources'
    ws2['A5'] = 'Sales'
    ws2['A6'] = 'Operations'
    ws2['A7'] = 'Legal'
    ws2['A8'] = 'Customer Support'
    ws2['A9'] = 'Research & Development'
    ws2['A10'] = 'Product Management'

    ws2.column_dimensions['A'].width = 25

    # No data validations in the initial file
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
