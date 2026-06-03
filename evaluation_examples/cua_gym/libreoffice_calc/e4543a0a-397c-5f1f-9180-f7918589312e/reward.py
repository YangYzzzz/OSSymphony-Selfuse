"""
Reward Script: Goal Seek to find Unit Price for $500,000 revenue target
Task ID: calc_gg5_021
Domain: libreoffice_calc
Scoring:
  - Gate: F3 must contain formula =B3*C3*D3*(1-E3) (precondition, no points)
  - Component 1 (0.6): B3 value is approximately 66.84 (Goal Seek result)
  - Component 2 (0.4): B3 changed from 45 AND C3, D3, E3 remain unchanged
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'sales_analysis'

def persist_app_state(domain: str):
    """Attempt to save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify Goal Seek task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb['Data']
    except KeyError:
        print("CRITICAL: 'Data' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: F3 must contain the revenue formula
    # This is true in both initial and golden, so it's a gate, not a scoring component.
    try:
        wb_formulas = openpyxl.load_workbook(file_path, data_only=False)
        ws_formulas = wb_formulas['Data']
        f3_val = ws_formulas['F3'].value
        expected_formula = "=B3*C3*D3*(1-E3)"
        if not isinstance(f3_val, str):
            print(f"FAIL: Precondition — F3 is not a formula: {f3_val}")
            print("REWARD: 0.0")
            return 0.0
        normalized = f3_val.upper().replace(" ", "")
        expected_normalized = expected_formula.upper().replace(" ", "")
        if normalized != expected_normalized:
            print(f"FAIL: Precondition — F3 formula = '{f3_val}', expected '{expected_formula}'")
            print("REWARD: 0.0")
            return 0.0
        print(f"GATE: F3 formula intact: '{f3_val}'")
    except Exception as e:
        print(f"ERROR: Precondition check — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: B3 value is approximately 66.84 (Goal Seek result for 500000 target) (0.6 points)
    # The exact value is 500000 / (8500 * 1 * (1-0.12)) = 500000 / 7480 = 66.84491978609626
    try:
        b3_val = ws['B3'].value
        expected_b3 = 66.84491978609626
        if b3_val is not None and isinstance(b3_val, (int, float)):
            if abs(float(b3_val) - expected_b3) < 0.1:
                print(f"PASS: Component 1 — B3 = {b3_val}, matches expected {expected_b3} (0.6 pts)")
                total_score += 0.6
            elif abs(float(b3_val) - expected_b3) < 1.0:
                print(f"PARTIAL: Component 1 — B3 = {b3_val}, close to expected {expected_b3} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 — B3 = {b3_val}, expected ~{expected_b3}")
        else:
            print(f"FAIL: Component 1 — B3 is not a numeric value: {b3_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 changed AND C3, D3, E3 remain unchanged (0.4 points)
    # This verifies Goal Seek only modified B3 (the variable cell) and left others intact
    try:
        b3_val = ws['B3'].value
        c3_val = ws['C3'].value
        d3_val = ws['D3'].value
        e3_val = ws['E3'].value

        b3_changed = False
        if b3_val is not None and isinstance(b3_val, (int, float)):
            if abs(float(b3_val) - 45.0) > 1.0:
                b3_changed = True

        others_intact = True
        if c3_val is None or abs(float(c3_val) - 8500) >= 0.01:
            others_intact = False
            print(f"  INFO: C3 changed — got {c3_val}, expected 8500")
        if d3_val is None or abs(float(d3_val) - 1.0) >= 0.01:
            others_intact = False
            print(f"  INFO: D3 changed — got {d3_val}, expected 1.0")
        if e3_val is None or abs(float(e3_val) - 0.12) >= 0.001:
            others_intact = False
            print(f"  INFO: E3 changed — got {e3_val}, expected 0.12")

        if b3_changed and others_intact:
            print(f"PASS: Component 2 — B3 modified ({b3_val}), C3/D3/E3 unchanged (0.4 pts)")
            total_score += 0.4
        elif not b3_changed:
            print(f"FAIL: Component 2 — B3 not modified (still {b3_val})")
        else:
            print(f"FAIL: Component 2 — Other cells were unexpectedly changed")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
