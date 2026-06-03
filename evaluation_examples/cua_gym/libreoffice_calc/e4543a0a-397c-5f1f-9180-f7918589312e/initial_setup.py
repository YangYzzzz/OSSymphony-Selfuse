"""
Initial Setup: Goal Seek to find unit price for target revenue
Task ID: calc_gg5_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'sales_analysis'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Data ---
    ws = wb.active
    ws.title = 'Data'

    # Header row styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Headers
    headers = ['Product', 'Unit Price', 'Units Sold', 'Upsell Factor', 'Discount Rate', 'Total Revenue']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows - realistic product sales data
    # Row 2
    products = [
        ['Standard Widget',      32.50,  12000, 1.0,  0.08],
        ['Premium Widget',       45.00,   8500, 1.0,  0.12],
        ['Deluxe Widget',        78.25,   4200, 1.15, 0.10],
        ['Economy Pack',         18.00,  25000, 1.0,  0.15],
        ['Professional Suite',  125.00,   2800, 1.20, 0.05],
        ['Enterprise License',  250.00,   1100, 1.10, 0.03],
        ['Starter Kit',          22.75,  18500, 1.0,  0.10],
        ['Maintenance Plan',     55.00,   6700, 1.05, 0.07],
        ['Training Bundle',      89.99,   3400, 1.0,  0.12],
        ['Custom Solution',     175.00,   1900, 1.25, 0.02],
        ['Volume License',       15.50,  42000, 1.0,  0.18],
    ]

    currency_fmt = '$#,##0.00'
    pct_fmt = '0%'
    number_fmt = '#,##0'
    revenue_fmt = '$#,##0.00'

    for r, row_data in enumerate(products, 2):
        name, price, units, upsell, discount = row_data
        ws.cell(row=r, column=1, value=name).border = thin_border
        c_price = ws.cell(row=r, column=2, value=price)
        c_price.number_format = currency_fmt
        c_price.border = thin_border
        c_units = ws.cell(row=r, column=3, value=units)
        c_units.number_format = number_fmt
        c_units.border = thin_border
        c_upsell = ws.cell(row=r, column=4, value=upsell)
        c_upsell.number_format = '0.00'
        c_upsell.border = thin_border
        c_discount = ws.cell(row=r, column=5, value=discount)
        c_discount.number_format = pct_fmt
        c_discount.border = thin_border
        # Formula for Total Revenue: Unit Price * Units Sold * Upsell Factor * (1 - Discount Rate)
        c_rev = ws.cell(row=r, column=6, value=f'=B{r}*C{r}*D{r}*(1-E{r})')
        c_rev.number_format = revenue_fmt
        c_rev.border = thin_border

    # Align data cells
    data_align = Alignment(horizontal="right", vertical="center")
    name_align = Alignment(horizontal="left", vertical="center")
    for r in range(2, 2 + len(products)):
        ws.cell(row=r, column=1).alignment = name_align
        for c in range(2, 7):
            ws.cell(row=r, column=c).alignment = data_align

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Sales Analysis Summary'
    ws2['A1'].font = Font(name="Calibri", size=14, bold=True)
    ws2['A3'] = 'Report Date:'
    ws2['B3'] = '2025-09-15'
    ws2['A4'] = 'Prepared By:'
    ws2['B4'] = 'Finance Department'
    ws2['A6'] = 'Total Products:'
    ws2['B6'] = len(products)
    ws2['A7'] = 'Revenue Target:'
    ws2['B7'] = 500000
    ws2['B7'].number_format = '$#,##0'
    ws2['A8'] = 'Notes:'
    ws2['B8'] = 'Use Goal Seek on Premium Widget to determine required price for target revenue.'
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 55

    # Make Data the active sheet
    wb.active = wb.sheetnames.index('Data')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
