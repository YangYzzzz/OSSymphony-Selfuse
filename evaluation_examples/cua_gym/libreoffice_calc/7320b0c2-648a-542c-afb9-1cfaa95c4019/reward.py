"""
Reward Script: Vendor Payment Data Cleanup
Task ID: calc_gen_data_cleanup_041
Domain: libreoffice_calc

Task: Consolidate vendor payment data by:
  1. Adding an 'Amount' column (G) with IF formula to consolidate System1/System2 amounts
  2. Removing duplicate Payment IDs (keep first occurrence)
  3. Adding conditional formatting: red fill on G when Amount > 10000 AND not 'Yes' in Approved

Scoring Rubric:
  Component 1: Column G header is 'Amount'                             — 0.20 points
  Component 2: Column G has IF formulas consolidating D and E          — 0.30 points
  Component 3: Duplicate Payment IDs removed (all IDs unique)          — 0.30 points
  Component 4: Conditional formatting applied on column G              — 0.20 points
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_041'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load file — precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Payments' sheet must exist
    if 'Payments' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Payments' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Payments']

    # -----------------------------------------------------------------------
    # Component 1: Column G header is 'Amount' (0.20 points)
    # This FAILS on initial (G1 is None) and PASSES on golden (G1 = 'Amount')
    # -----------------------------------------------------------------------
    try:
        g_header = ws.cell(row=1, column=7).value
        if g_header and str(g_header).strip() == 'Amount':
            print(f"PASS: Component 1 — Column G header is 'Amount' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Expected G1='Amount', found: {repr(g_header)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Column G has IF formulas consolidating D and E (0.30 points)
    # The formula pattern should be =IF(D<n><>"",D<n>,E<n>) for each data row.
    # This FAILS on initial (G2:G201 are all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        max_row = ws.max_row
        if max_row < 2:
            print("FAIL: Component 2 — No data rows found")
        else:
            # Count rows with correct IF formula pattern
            formula_pattern = re.compile(
                r'^=IF\(D\d+<>"",D\d+,E\d+\)$',
                re.IGNORECASE
            )
            formula_count = 0
            total_data_rows = max_row - 1  # exclude header
            non_formula_rows = []

            for row in range(2, max_row + 1):
                cell_val = ws.cell(row=row, column=7).value
                if cell_val is None:
                    non_formula_rows.append((row, None))
                elif isinstance(cell_val, str) and formula_pattern.match(cell_val.strip()):
                    formula_count += 1
                else:
                    non_formula_rows.append((row, cell_val))

            if non_formula_rows:
                for bad_row, bad_val in non_formula_rows[:3]:
                    print(f"  Note: Row {bad_row} G = {repr(bad_val)}")

            ratio = formula_count / total_data_rows if total_data_rows > 0 else 0
            if ratio >= 0.95:  # allow small tolerance
                print(f"PASS: Component 2 — Column G IF formulas: {formula_count}/{total_data_rows} rows (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 2 — Only {formula_count}/{total_data_rows} rows have IF formula in G")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Duplicate Payment IDs removed — all IDs unique (0.30 points)
    # Initial file has 200 rows with ~14 duplicates; golden should have 186 unique.
    # This FAILS on initial (200 rows, duplicates present) and PASSES on golden (186 rows, no duplicates).
    # -----------------------------------------------------------------------
    try:
        payment_ids = []
        for row in range(2, ws.max_row + 1):
            pid = ws.cell(row=row, column=1).value
            if pid is not None:
                payment_ids.append(pid)

        total_ids = len(payment_ids)
        unique_ids = len(set(payment_ids))
        duplicate_count = total_ids - unique_ids

        if duplicate_count == 0 and total_ids > 0:
            print(f"PASS: Component 3 — No duplicate Payment IDs: {unique_ids} unique rows (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — Found {duplicate_count} duplicate Payment IDs "
                  f"({total_ids} total, {unique_ids} unique)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional formatting on column G (0.20 points)
    # Red fill rule: AND(G>10000, F<>"Yes") applied to G2:G<last>
    # This FAILS on initial (no CF on G) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        cf_found = False
        cf_details = []

        for cf_range in cf_rules:
            range_str = str(cf_range)
            # Check if this CF applies to column G
            if 'G' in range_str:
                for rule in cf_rules[cf_range]:
                    rule_type = getattr(rule, 'type', '')
                    formula = getattr(rule, 'formula', [])
                    formula_str = str(formula[0]).upper() if formula else ''

                    # Check for formula containing G>10000 and F<>"Yes" logic
                    has_amount_check = '10000' in formula_str
                    has_approved_check = 'YES' in formula_str or '"YES"' in formula_str

                    # Check for red fill
                    has_red_fill = False
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            # Red fill: starts with FF (opaque) and has high R, low G/B
                            # Accept FFFF0000 (pure red) or similar red variants
                            if fill_color and fill_color.upper().startswith('FF') and fill_color.upper() in (
                                'FFFF0000', 'FFFF4444', 'FFCC0000', 'FFDC143C',
                                'FFFF3333', 'FFFF6666', 'FFFFE0E0', 'FFFFC7CE'
                            ):
                                has_red_fill = True
                            elif fill_color:
                                # Check if it's any red-ish fill (R component high, others low)
                                # ARGB: FF RRRR GGGG BBBB
                                try:
                                    r_val = int(fill_color[2:4], 16)
                                    g_val = int(fill_color[4:6], 16)
                                    b_val = int(fill_color[6:8], 16)
                                    if r_val > 150 and g_val < 100 and b_val < 100:
                                        has_red_fill = True
                                    # Also accept light pink/salmon often used for Excel conditional formatting
                                    elif r_val > 200 and g_val > 150 and b_val > 150:
                                        has_red_fill = True  # light red
                                except:
                                    pass
                        except Exception as fill_e:
                            print(f"  Fill check error: {fill_e}")

                    cf_details.append({
                        'range': range_str,
                        'type': rule_type,
                        'formula': formula_str,
                        'has_amount_check': has_amount_check,
                        'has_approved_check': has_approved_check,
                        'has_red_fill': has_red_fill,
                    })

                    if has_amount_check and has_approved_check and has_red_fill:
                        cf_found = True

        if cf_found:
            print(f"PASS: Component 4 — Conditional formatting on G with red fill "
                  f"for Amount>10000 AND not Approved (0.20 pts)")
            total_score += 0.20
        else:
            # Partial check: CF exists on G column at all
            g_cf_exists = any('G' in d['range'] for d in cf_details)
            if cf_details:
                print(f"FAIL: Component 4 — CF on G column found but conditions not fully met:")
                for d in cf_details:
                    print(f"  range={d['range']}, formula={d['formula']}, "
                          f"amount_check={d['has_amount_check']}, "
                          f"approved_check={d['has_approved_check']}, "
                          f"red_fill={d['has_red_fill']}")
            else:
                print(f"FAIL: Component 4 — No conditional formatting found on column G")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
