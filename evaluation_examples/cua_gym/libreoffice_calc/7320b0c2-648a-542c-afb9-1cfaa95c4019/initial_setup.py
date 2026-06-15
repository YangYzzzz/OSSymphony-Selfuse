"""
Initial Setup: Vendor payment data cleanup — split amounts, duplicate IDs
Task ID: calc_gen_data_cleanup_041
Domain: libreoffice_calc

Creates a 'Payments' sheet with 200 rows of vendor payment data.
- Columns D (Amount System1) and E (Amount System2) are mutually exclusive
- ~15 duplicate Payment IDs exist
- Approved (F): 'Yes', 'No', or blank
- Column G is empty (no Amount column yet)
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payments'

    # Headers
    headers = ['Payment ID', 'Vendor', 'Date', 'Amount System1', 'Amount System2', 'Approved']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    # Column G intentionally empty (no header)

    # Realistic vendor names
    vendors = [
        'Acme Supplies Co.', 'Global Tech Partners', 'Pacific Rim Logistics',
        'Northbridge Consulting', 'Redwood Materials', 'Summit Office Solutions',
        'Horizon Digital Services', 'BlueSky Analytics', 'Pinnacle Hardware',
        'Meridian Freight LLC', 'Cascade Printing Group', 'Ironwood Fabrication',
        'Coastal Data Systems', 'Oakdale Catering', 'Sterling Software Inc.',
        'Apex Electrical', 'Greenfield Landscaping', 'Landmark Security',
        'Vanguard Staffing', 'Metro Cleaning Services'
    ]

    # Approved status distribution
    approved_values = ['Yes', 'No', '']
    approved_dist = [0, 0, 1, 0, 1, 0, 0, 1, 0, 1]  # used cyclically

    # Unique Payment IDs base (we'll insert ~15 duplicates into the 200 rows)
    # Generate unique IDs first
    unique_ids = [f'PAY-{2025000 + i:07d}' for i in range(1, 187)]  # 186 unique IDs

    # 15 duplicate IDs to scatter through the data (duplicate of some early IDs)
    dup_ids = [f'PAY-{2025000 + i:07d}' for i in range(1, 16)]  # IDs 1-15 will appear twice

    # Build a list of 200 Payment IDs (186 unique + 14 duplicates scattered in)
    # Insert duplicate IDs at specific positions
    all_ids = list(unique_ids)  # 186 unique
    # Insert 14 duplicates at evenly spaced positions across 200 rows
    dup_positions = [20, 35, 50, 68, 80, 95, 110, 125, 138, 152, 163, 175, 185, 195]
    for pos, dup_id in zip(dup_positions, dup_ids[:14]):
        all_ids.insert(pos, dup_id)
    # Now all_ids has 200 entries

    # Amount values — some realistic vendor payment amounts
    amounts_low = [
        1250.00, 3480.75, 875.50, 6200.00, 15750.00, 490.25, 22000.00, 8900.00,
        1100.00, 5600.00, 340.00, 12500.00, 760.00, 4300.00, 9800.00, 2100.00,
        45000.00, 3750.00, 620.00, 18200.00, 1950.00, 7400.00, 280.00, 11000.00,
        5100.00, 14600.00, 830.00, 3200.00, 9000.00, 2700.00, 6800.00, 1600.00,
        31000.00, 430.00, 8200.00, 1750.00, 25000.00, 4600.00, 960.00, 7100.00
    ]
    amounts_high = [
        5500.00, 13200.00, 2800.00, 7600.00, 1400.00, 19500.00, 3300.00, 900.00,
        22500.00, 6100.00, 480.00, 16800.00, 4200.00, 8100.00, 2350.00, 10500.00,
        630.00, 7900.00, 3600.00, 14000.00, 1800.00, 5700.00, 11500.00, 2100.00,
        750.00, 8600.00, 4100.00, 20000.00, 1250.00, 6400.00, 3900.00, 9300.00,
        550.00, 17500.00, 2950.00, 12000.00, 4800.00, 1100.00, 6700.00, 28000.00
    ]

    # Date values
    date_list = [
        '2025-01-03', '2025-01-07', '2025-01-10', '2025-01-14', '2025-01-17',
        '2025-01-21', '2025-01-24', '2025-01-28', '2025-02-03', '2025-02-06',
        '2025-02-10', '2025-02-13', '2025-02-17', '2025-02-20', '2025-02-24',
        '2025-02-27', '2025-03-03', '2025-03-06', '2025-03-10', '2025-03-13',
        '2025-03-17', '2025-03-20', '2025-03-24', '2025-03-27', '2025-03-31',
        '2025-04-03', '2025-04-07', '2025-04-10', '2025-04-14', '2025-04-17'
    ]

    for i, pay_id in enumerate(all_ids):
        row = i + 2
        vendor = vendors[i % len(vendors)]
        date = date_list[i % len(date_list)]
        # Mutually exclusive amounts: even rows use System1, odd rows use System2
        amt_index = i % len(amounts_low)
        if i % 2 == 0:
            # System1 has value, System2 blank
            amt_d = amounts_low[amt_index]
            amt_e = None
        else:
            # System2 has value, System1 blank
            amt_d = None
            amt_e = amounts_high[amt_index]

        # Approved status
        if i % 3 == 0:
            approved = 'Yes'
        elif i % 3 == 1:
            approved = 'No'
        else:
            approved = ''

        ws.cell(row=row, column=1, value=pay_id)
        ws.cell(row=row, column=2, value=vendor)
        ws.cell(row=row, column=3, value=date)
        ws.cell(row=row, column=4, value=amt_d)
        ws.cell(row=row, column=5, value=amt_e)
        ws.cell(row=row, column=6, value=approved)
        # Column G intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Rows of data: {len(all_ids)} (headers + 200 data rows = 201 total)')
    print(f'Duplicate IDs inserted at positions: {dup_positions}')

create_initial()
