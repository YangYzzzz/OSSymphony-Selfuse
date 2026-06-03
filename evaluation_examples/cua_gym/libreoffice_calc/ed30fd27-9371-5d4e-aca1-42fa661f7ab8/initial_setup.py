"""
Initial Setup: Create regional sales data spreadsheet for territory chart task
Task ID: calc_sales_territory_chart_011
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'calc_sales_territory_chart_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: RegionalData ---
    ws = wb.active
    ws.title = 'RegionalData'

    # Headers in A1:E1
    headers = ['Region', 'Total Revenue', 'Quota', 'Attainment %', 'Deals Won']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows A2:E5
    # Region, Total Revenue, Quota, Attainment % (formula), Deals Won
    regions = ['North', 'South', 'East', 'West']
    revenues = [1800000, 2400000, 1200000, 3100000]
    quotas   = [2000000, 2200000, 1500000, 2800000]
    deals    = [45, 62, 31, 78]

    thin = Side(style='thin', color='FFB8B8B8')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, (region, rev, quota, deal) in enumerate(zip(regions, revenues, quotas, deals)):
        row = i + 2
        # Region name
        cell_a = ws.cell(row=row, column=1, value=region)
        cell_a.font = Font(name='Calibri', size=11, bold=True)
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        cell_a.border = border

        # Total Revenue
        cell_b = ws.cell(row=row, column=2, value=rev)
        cell_b.number_format = '$#,##0'
        cell_b.alignment = Alignment(horizontal='right', vertical='center')
        cell_b.border = border

        # Quota
        cell_c = ws.cell(row=row, column=3, value=quota)
        cell_c.number_format = '$#,##0'
        cell_c.alignment = Alignment(horizontal='right', vertical='center')
        cell_c.border = border

        # Attainment % — formula =B/C
        formula = f'=B{row}/C{row}'
        cell_d = ws.cell(row=row, column=4, value=formula)
        cell_d.number_format = '0.00%'
        cell_d.alignment = Alignment(horizontal='right', vertical='center')
        cell_d.border = border

        # Deals Won
        cell_e = ws.cell(row=row, column=5, value=deal)
        cell_e.alignment = Alignment(horizontal='center', vertical='center')
        cell_e.border = border

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12

    # Row heights
    ws.row_dimensions[1].height = 22
    for row in range(2, 6):
        ws.row_dimensions[row].height = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    # NO CHARTS — task requires creating the chart
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: RegionalData')
    print('Data: 4 regions with Revenue, Quota, Attainment %, Deals Won')
    print('No charts — task requires creating combo chart')


create_initial()
