"""
Reward Script: Create a regional performance comparison combo chart
Task ID: calc_sales_territory_chart_011
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: A chart exists on sheet 'RegionalData' (or a new chart sheet)   — 0.25 pts
  Component 2: Chart is a combo chart (both barChart and lineChart in plotArea)  — 0.25 pts
  Component 3: Chart title matches 'Regional Performance vs Quota'               — 0.20 pts
  Component 4: Bar series references Total Revenue data (column B)               — 0.15 pts
  Component 5: Line series references Attainment % data (column D)               — 0.15 pts
  Total: 1.0

Strategy:
  - Parse the chart XML directly from the xlsx zip to reliably detect combo charts
    (openpyxl's high-level API only exposes the first chart type in a combo chart).
  - Title, series data references, and axis titles are extracted from raw XML.
"""

import os
import zipfile
import xml.etree.ElementTree as ET

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_territory_chart_011'

# XML namespaces used in xlsx chart files
NS_C = 'http://schemas.openxmlformats.org/drawingml/2006/chart'
NS_A = 'http://schemas.openxmlformats.org/drawingml/2006/main'


def get_text_from_title_element(title_elem):
    """Extract plain text from a chart title XML element."""
    if title_elem is None:
        return None
    texts = []
    for t in title_elem.iter(f'{{{NS_A}}}t'):
        if t.text:
            texts.append(t.text.strip())
    return ' '.join(texts) if texts else None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # --- Precondition: file must be openable ---
    if not os.path.exists(file_path):
        print(f"CRITICAL: File not found: {file_path}")
        print("REWARD: 0.0")
        return 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition: RegionalData sheet must exist ---
    if 'RegionalData' not in wb.sheetnames:
        print("CRITICAL: Sheet 'RegionalData' not found")
        print("REWARD: 0.0")
        return 0.0

    # --- Parse chart XML directly from the zip ---
    chart_xmls = []
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            chart_files = [n for n in z.namelist() if 'charts/chart' in n.lower()]
            for cf in chart_files:
                content = z.read(cf).decode('utf-8')
                root = ET.fromstring(content)
                chart_xmls.append(root)
    except Exception as e:
        print(f"ERROR: Could not read chart XML from zip: {e}")

    # --- Component 1: At least one chart exists on RegionalData or any sheet (0.25 pts) ---
    # Check via openpyxl (charts list) plus XML presence
    try:
        chart_found = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if len(ws._charts) > 0:
                chart_found = True
                break
        # Also consider: chart XML files present
        if not chart_found and len(chart_xmls) > 0:
            chart_found = True

        if chart_found:
            print("PASS: Component 1 — Chart exists in the workbook (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 1 — No chart found in the workbook")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: Combo chart — both barChart and lineChart present (0.25 pts) ---
    try:
        bar_found = False
        line_found = False

        for root in chart_xmls:
            bar_elems = root.findall(f'.//{{{NS_C}}}barChart')
            line_elems = root.findall(f'.//{{{NS_C}}}lineChart')
            if bar_elems:
                bar_found = True
            if line_elems:
                line_found = True

        if bar_found and line_found:
            print("PASS: Component 2 — Combo chart with both barChart and lineChart (0.25 pts)")
            total_score += 0.25
        elif bar_found:
            print("FAIL: Component 2 — Only barChart found; no lineChart (Attainment % line missing)")
        elif line_found:
            print("FAIL: Component 2 — Only lineChart found; no barChart (Revenue bars missing)")
        else:
            print("FAIL: Component 2 — Neither barChart nor lineChart found in chart XML")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Chart title is 'Regional Performance vs Quota' (0.20 pts) ---
    try:
        title_matched = False
        expected_title = 'Regional Performance vs Quota'

        for root in chart_xmls:
            # The top-level title is under <chart><title>
            chart_elem = root.find(f'{{{NS_C}}}chart')
            if chart_elem is not None:
                title_elem = chart_elem.find(f'{{{NS_C}}}title')
                title_text = get_text_from_title_element(title_elem)
                if title_text:
                    if title_text.strip().lower() == expected_title.lower():
                        title_matched = True
                        print(f"PASS: Component 3 — Chart title is '{title_text}' (0.20 pts)")
                    else:
                        print(f"FAIL: Component 3 — Chart title is '{title_text}', expected '{expected_title}'")

        if title_matched:
            total_score += 0.20
        elif not chart_xmls:
            print(f"FAIL: Component 3 — No chart XML found to check title")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # --- Component 4: Bar series references Total Revenue (column B, rows 2-5) (0.15 pts) ---
    try:
        bar_revenue_ok = False
        for root in chart_xmls:
            bar_charts = root.findall(f'.//{{{NS_C}}}barChart')
            for bc in bar_charts:
                for ser in bc.findall(f'{{{NS_C}}}ser'):
                    val_elem = ser.find(f'.//{{{NS_C}}}val')
                    if val_elem is not None:
                        num_ref = val_elem.find(f'{{{NS_C}}}numRef')
                        if num_ref is not None:
                            f_elem = num_ref.find(f'{{{NS_C}}}f')
                            if f_elem is not None and f_elem.text:
                                ref = f_elem.text.strip()
                                # Accept any reference to column B rows 2-5
                                if 'B' in ref and ('B2' in ref or '$B$2' in ref):
                                    bar_revenue_ok = True
                                    print(f"PASS: Component 4 — Bar series references Revenue data ({ref}) (0.15 pts)")
                                    break
                    if bar_revenue_ok:
                        break

        if bar_revenue_ok:
            total_score += 0.15
        else:
            print("FAIL: Component 4 — Bar series does not reference column B (Total Revenue) data")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # --- Component 5: Line series references Attainment % (column D, rows 2-5) (0.15 pts) ---
    try:
        line_attainment_ok = False
        for root in chart_xmls:
            line_charts = root.findall(f'.//{{{NS_C}}}lineChart')
            for lc in line_charts:
                for ser in lc.findall(f'{{{NS_C}}}ser'):
                    val_elem = ser.find(f'.//{{{NS_C}}}val')
                    if val_elem is not None:
                        num_ref = val_elem.find(f'{{{NS_C}}}numRef')
                        if num_ref is not None:
                            f_elem = num_ref.find(f'{{{NS_C}}}f')
                            if f_elem is not None and f_elem.text:
                                ref = f_elem.text.strip()
                                # Accept any reference to column D rows 2-5
                                if 'D' in ref and ('D2' in ref or '$D$2' in ref):
                                    line_attainment_ok = True
                                    print(f"PASS: Component 5 — Line series references Attainment % data ({ref}) (0.15 pts)")
                                    break
                    if line_attainment_ok:
                        break

        if line_attainment_ok:
            total_score += 0.15
        else:
            print("FAIL: Component 5 — Line series does not reference column D (Attainment %) data")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
