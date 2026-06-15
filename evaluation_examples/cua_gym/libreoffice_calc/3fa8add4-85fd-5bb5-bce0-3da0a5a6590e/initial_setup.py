"""
Initial Setup: Create a spreadsheet with quarterly sales data (no charts)
Task ID: calc_chart_copy_format_061
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_copy_format_061'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DarkTheme ---
    ws = wb.active
    ws.title = 'DarkTheme'

    # Headers
    headers = ['Category', 'This Year', 'Last Year']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows (Q1-Q4) — exact values from task context
    data = [
        ['Q1', 148000, 132000],
        ['Q2', 165000, 148000],
        ['Q3', 172000, 155000],
        ['Q4', 198000, 168000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    # NO charts in initial file

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
