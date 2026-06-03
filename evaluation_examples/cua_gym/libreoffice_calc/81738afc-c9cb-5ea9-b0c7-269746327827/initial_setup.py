"""
Initial Setup: Create student grades spreadsheet for pivot table analysis
Task ID: calc_pivot_087
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def get_grade_letter(score):
    if score >= 90:
        return 'A'
    elif score >= 80:
        return 'B'
    elif score >= 70:
        return 'C'
    elif score >= 60:
        return 'D'
    else:
        return 'F'


def create_initial():
    random.seed(42)

    courses = ['Intro CS', 'Data Structures', 'Algorithms', 'Databases', 'Networks']

    # We need 250 students total, with specific constraints:
    # - Intro CS / A = 12
    # - Algorithms / F = 8
    # We'll distribute 50 students per course for simplicity.

    first_names = [
        'Sarah', 'Marcus', 'Emily', 'James', 'Olivia', 'Liam', 'Sophia', 'Noah',
        'Ava', 'Ethan', 'Mia', 'Lucas', 'Isabella', 'Mason', 'Charlotte',
        'Logan', 'Amelia', 'Alexander', 'Harper', 'Benjamin', 'Ella', 'Daniel',
        'Aria', 'Henry', 'Chloe', 'Jackson', 'Luna', 'Sebastian', 'Grace', 'Aiden',
        'Lily', 'Owen', 'Zoey', 'Samuel', 'Nora', 'Ryan', 'Riley', 'Nathan',
        'Scarlett', 'Caleb', 'Hannah', 'Dylan', 'Layla', 'Isaac', 'Penelope',
        'Wyatt', 'Eleanor', 'Gabriel', 'Violet', 'Julian', 'Hazel', 'Leo',
        'Stella', 'Jayden', 'Aurora', 'Carter', 'Savannah', 'Luke', 'Audrey',
        'Jack', 'Brooklyn', 'Matthew', 'Bella', 'Adam', 'Claire', 'Asher',
        'Skylar', 'Thomas', 'Lucy', 'Hudson', 'Paisley', 'Connor', 'Everly',
        'Eli', 'Anna', 'Lincoln', 'Caroline', 'Jose', 'Kennedy', 'Jaxon',
        'Maya', 'Dominic', 'Willow', 'Austin', 'Kinsley', 'Ian', 'Naomi',
        'Cooper', 'Aaliyah', 'Robert', 'Elena', 'Easton', 'Eliana', 'Colton',
        'Peyton', 'Jordan', 'Madeline', 'Carson', 'Ruby', 'Cameron', 'Eva'
    ]
    last_names = [
        'Chen', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
        'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
        'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
        'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
        'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
        'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green',
        'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts', 'Kim', 'Park', 'Patel', 'Shah', 'Singh',
        'Kumar', 'Sharma', 'Gupta', 'Li', 'Wang', 'Zhang', 'Liu', 'Yang',
        'Huang', 'Wu', 'Zhou', 'Xu', 'Sun', 'Ma', 'Zhu'
    ]

    # Generate names (cycle through combinations)
    names = []
    for i in range(250):
        fn = first_names[i % len(first_names)]
        ln = last_names[i % len(last_names)]
        names.append(f'{fn} {ln}')

    # Generate scores per course with constraints
    # 50 students per course
    students_per_course = 50
    all_students = []
    student_id = 1

    for course in courses:
        scores = []
        if course == 'Intro CS':
            # Need exactly 12 students with A (90-100)
            a_scores = [random.randint(90, 100) for _ in range(12)]
            # Fill remaining 38 with B, C, D, F range (25-89)
            other_scores = [random.randint(25, 89) for _ in range(38)]
            scores = a_scores + other_scores
            random.shuffle(scores)
        elif course == 'Algorithms':
            # Need exactly 8 students with F (below 60)
            f_scores = [random.randint(25, 59) for _ in range(8)]
            # Fill remaining 42 with scores 60-100
            other_scores = [random.randint(60, 100) for _ in range(42)]
            scores = f_scores + other_scores
            random.shuffle(scores)
        else:
            # Random distribution for other courses
            scores = [random.randint(25, 100) for _ in range(students_per_course)]

        for score in scores:
            grade = get_grade_letter(score)
            all_students.append((student_id, names[student_id - 1], course, score, grade))
            student_id += 1

    # Shuffle all students so they're not grouped by course
    random.shuffle(all_students)
    # Reassign student IDs after shuffle
    final_students = []
    for idx, (_, name, course, score, grade) in enumerate(all_students, 1):
        final_students.append((idx, name, course, score, grade))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Grades'

    # Headers with formatting
    headers = ['StudentID', 'Name', 'Course', 'FinalScore', 'GradeLetter']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    thin_border = Border(
        bottom=Side(style="thin", color="000000")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data
    for r, (sid, name, course, score, grade) in enumerate(final_students, 2):
        ws.cell(row=r, column=1, value=sid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=course)
        ws.cell(row=r, column=4, value=score)
        ws.cell(row=r, column=5, value=grade)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify counts
    from collections import Counter
    course_grade_counts = Counter()
    for _, _, course, _, grade in final_students:
        course_grade_counts[(course, grade)] += 1
    print(f"Intro CS / A = {course_grade_counts[('Intro CS', 'A')]}")
    print(f"Algorithms / F = {course_grade_counts[('Algorithms', 'F')]}")
    print(f"Total students = {len(final_students)}")

    # GUI launch
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
