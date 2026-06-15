"""
Reward Script: Create a pivot table analyzing student performance by grade band and course.
Task ID: calc_pivot_087
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Pivot sheet exists (new sheet created for the pivot table)
  Component 2 (0.25): Correct structure — Course as rows, grade letters A/B/C/D/F as columns
  Component 3 (0.25): Correct data values for specific known cells (Intro CS/A=12, Algorithms/F=8)
  Component 4 (0.15): Grand Total row and column with correct totals (250)
  Component 5 (0.15): Columns ordered A, B, C, D, F
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_087'

# Expected courses and grade columns from the task context
EXPECTED_COURSES = ['Algorithms', 'Data Structures', 'Databases', 'Intro CS', 'Networks']
EXPECTED_GRADE_ORDER = ['A', 'B', 'C', 'D', 'F']


def persist_app_state(domain: str):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def find_pivot_sheet(wb):
    """Find a sheet that looks like the pivot table (not 'Grades')."""
    for sn in wb.sheetnames:
        if sn.lower() != 'grades':
            ws = wb[sn]
            # Check if it has content that looks like a pivot table
            if ws.max_row >= 2 and ws.max_column >= 2:
                return ws
    return None


def read_pivot_data(ws):
    """Read the pivot table into a dict structure.
    Returns: (headers_list, data_dict, header_row_idx, data_start_row, data_end_row)
    where data_dict[course] = {grade: count, ...}
    """
    # Find header row (row that contains grade letters)
    header_row = None
    headers = []
    for r in range(1, min(ws.max_row + 1, 10)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        # Check if this row contains multiple grade letters
        grade_count = sum(1 for v in row_vals if v in ('A', 'B', 'C', 'D', 'F'))
        if grade_count >= 3:
            header_row = r
            headers = row_vals
            break

    if header_row is None:
        return None, None, None, None, None

    # Read data rows
    data_dict = {}
    data_start = header_row + 1
    data_end = ws.max_row
    for r in range(data_start, data_end + 1):
        course = ws.cell(row=r, column=1).value
        if course is None:
            continue
        course_str = str(course).strip()
        row_data = {}
        for c in range(2, len(headers) + 1):
            h = headers[c - 1]
            if h is not None:
                row_data[str(h).strip()] = ws.cell(row=r, column=c).value
        data_dict[course_str] = row_data

    return headers, data_dict, header_row, data_start, data_end


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot sheet exists (0.20 points)
    # The initial file only has 'Grades'. A new sheet must be created.
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_ws.title}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — No pivot/summary sheet found besides 'Grades'. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_ws is None:
        # Can't check further without a pivot sheet
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Parse the pivot table
    headers, data_dict, header_row, data_start, data_end = read_pivot_data(pivot_ws)

    if headers is None or data_dict is None:
        print(f"FAIL: Could not parse pivot table structure in sheet '{pivot_ws.title}'")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Component 2: Correct structure — courses as rows, grades as columns (0.25 points)
    # Need all 5 courses present and grade letter columns
    try:
        # Check courses present (case-insensitive matching)
        found_courses = [k for k in data_dict.keys() if k in EXPECTED_COURSES]
        grade_cols = [str(h).strip() for h in headers[1:] if h is not None and str(h).strip() in ('A', 'B', 'C', 'D', 'F')]

        courses_ok = len(found_courses) == 5
        grades_ok = len(grade_cols) == 5

        if courses_ok and grades_ok:
            print(f"PASS: Component 2 — All 5 courses and 5 grade columns found (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Found {len(found_courses)}/5 courses, {len(grade_cols)}/5 grade columns")
            print(f"  Courses found: {found_courses}")
            print(f"  Grade columns: {grade_cols}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct data values for known cells (0.25 points)
    # From context: Intro CS/A=12, Algorithms/F=8
    try:
        checks_passed = 0
        total_checks = 2

        # Check Intro CS / A = 12
        intro_cs_data = data_dict.get('Intro CS', {})
        intro_cs_a = intro_cs_data.get('A')
        if intro_cs_a is not None and int(intro_cs_a) == 12:
            print(f"  PASS: Intro CS / A = {intro_cs_a} (expected 12)")
            checks_passed += 1
        else:
            print(f"  FAIL: Intro CS / A = {intro_cs_a} (expected 12)")

        # Check Algorithms / F = 8
        algo_data = data_dict.get('Algorithms', {})
        algo_f = algo_data.get('F')
        if algo_f is not None and int(algo_f) == 8:
            print(f"  PASS: Algorithms / F = {algo_f} (expected 8)")
            checks_passed += 1
        else:
            print(f"  FAIL: Algorithms / F = {algo_f} (expected 8)")

        if checks_passed == total_checks:
            print(f"PASS: Component 3 — Both known values correct (0.25 pts)")
            total_score += 0.25
        elif checks_passed > 0:
            partial = round(0.25 * checks_passed / total_checks, 3)
            print(f"PARTIAL: Component 3 — {checks_passed}/{total_checks} values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No known values matched")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row and column with correct totals (0.15 points)
    # Grand total should be 250 (total students)
    try:
        # Look for a row labeled 'Grand Total' or similar
        grand_total_row = data_dict.get('Grand Total', {})
        if not grand_total_row:
            # Try case-insensitive
            for k, v in data_dict.items():
                if 'total' in str(k).lower() and 'grand' in str(k).lower():
                    grand_total_row = v
                    break

        # Check if a Grand Total column exists in headers
        gt_col_idx = None
        for i, h in enumerate(headers):
            if h is not None and 'grand' in str(h).lower() and 'total' in str(h).lower():
                gt_col_idx = i
                break

        # Check grand total value = 250
        gt_value = None
        if grand_total_row:
            # Sum of grade columns in grand total row, or check Grand Total column
            if 'Grand Total' in grand_total_row:
                gt_value = grand_total_row['Grand Total']
            elif gt_col_idx is not None:
                # Get from the header index
                gt_key = str(headers[gt_col_idx]).strip()
                gt_value = grand_total_row.get(gt_key)

            # If not found directly, sum the grade values
            if gt_value is None:
                grade_sum = 0
                for g in EXPECTED_GRADE_ORDER:
                    v = grand_total_row.get(g)
                    if v is not None:
                        grade_sum += int(v)
                if grade_sum > 0:
                    gt_value = grade_sum

        if gt_value is not None and int(gt_value) == 250:
            print(f"PASS: Component 4 — Grand Total = 250 found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Grand Total = 250 not found. gt_value={gt_value}, Grand total row: {grand_total_row}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Columns ordered A, B, C, D, F (0.15 points)
    try:
        # Extract just the grade letter headers in order
        grade_headers_in_order = []
        for h in headers[1:]:
            if h is not None:
                hs = str(h).strip()
                if hs in ('A', 'B', 'C', 'D', 'F'):
                    grade_headers_in_order.append(hs)

        if grade_headers_in_order == EXPECTED_GRADE_ORDER:
            print(f"PASS: Component 5 — Grade columns in correct order A,B,C,D,F (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — Grade column order is {grade_headers_in_order}, expected {EXPECTED_GRADE_ORDER}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
