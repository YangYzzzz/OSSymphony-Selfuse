"""
Reward Script: Hide all rows where the Score column contains 'N/A' or is blank
Task ID: osworld_calc_hide_rows_na_005
Domain: libreoffice_calc

Scoring:
  Component 1: All N/A and blank Score rows are hidden (rows 4, 7, 10, 14, 17) — 0.7 points
  Component 2: A count of hidden rows (5) is noted somewhere in the sheet — 0.3 points
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_hide_rows_na_005'

# Rows expected to be hidden: rows 4, 7, 10, 14, 17
# These correspond to respondents with Score='N/A' or Score=None (blank)
EXPECTED_HIDDEN_ROWS = [4, 7, 10, 14, 17]
EXPECTED_HIDDEN_COUNT = 5


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    Task: Hide all rows where Score is 'N/A' or blank (do not use AutoFilter).
    Also note the total count of hidden rows somewhere in the sheet.

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — if this fails, the file is broken
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Survey Results sheet
    try:
        if 'Survey Results' not in wb.sheetnames:
            print(f"FAIL: Sheet 'Survey Results' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0
        ws = wb['Survey Results']
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Verify the Score column exists (col 3)
    try:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        if 'Score' not in headers:
            print(f"FAIL: 'Score' header not found. Headers: {headers}")
            print("REWARD: 0.0")
            return 0.0
        score_col = headers.index('Score') + 1  # 1-based
    except Exception as e:
        print(f"CRITICAL: Cannot read headers: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: All N/A and blank Score rows are hidden (0.7 points)
    #
    # In the initial_env, rows 4, 7, 10, 14, 17 have N/A or blank scores and
    # are NOT hidden. After the task, all of these rows must be hidden.
    # We verify by checking row_dimensions[row_num].hidden == True for each
    # of the expected rows.
    # -----------------------------------------------------------------------
    try:
        # Find all data rows (skip header row 1) where Score is N/A or blank
        na_blank_rows = []
        for row_num in range(2, ws.max_row + 1):
            score_val = ws.cell(row=row_num, column=score_col).value
            if score_val is None or str(score_val).strip().upper() == 'N/A':
                na_blank_rows.append(row_num)

        print(f"Rows with N/A or blank Score: {na_blank_rows}")
        print(f"Expected rows to hide: {EXPECTED_HIDDEN_ROWS}")

        # Check that the expected rows are hidden
        all_expected_hidden = True
        for row_num in EXPECTED_HIDDEN_ROWS:
            is_hidden = ws.row_dimensions[row_num].hidden
            if not is_hidden:
                print(f"FAIL: Row {row_num} (Score={ws.cell(row=row_num, column=score_col).value!r}) is NOT hidden")
                all_expected_hidden = False
            else:
                print(f"PASS: Row {row_num} is correctly hidden")

        # Also check that visible data rows are NOT incorrectly hidden
        all_visible_rows_correct = True
        for row_num in range(2, ws.max_row + 1):
            if row_num not in EXPECTED_HIDDEN_ROWS:
                score_val = ws.cell(row=row_num, column=score_col).value
                is_hidden = ws.row_dimensions[row_num].hidden
                # We only flag visible-data rows that are hidden but shouldn't be
                # (rows with numeric scores should NOT be hidden)
                if is_hidden and score_val is not None and str(score_val).strip().upper() != 'N/A':
                    try:
                        float(score_val)
                        # It's a numeric score — should not be hidden
                        print(f"WARN: Row {row_num} has numeric Score={score_val!r} but is hidden (over-hiding)")
                        all_visible_rows_correct = False
                    except (ValueError, TypeError):
                        pass  # Non-numeric non-N/A, skip

        if all_expected_hidden and all_visible_rows_correct:
            print(f"PASS: Component 1 — All 5 N/A/blank score rows are correctly hidden (0.7 pts)")
            total_score += 0.7
        elif all_expected_hidden and not all_visible_rows_correct:
            # Partial: hid the right rows but also over-hid valid rows
            print(f"PARTIAL: Component 1 — Expected rows hidden but some valid rows over-hidden (0.4 pts)")
            total_score += 0.4
        else:
            # Count how many of the expected rows are hidden for partial credit
            hidden_count = sum(1 for r in EXPECTED_HIDDEN_ROWS if ws.row_dimensions[r].hidden)
            if hidden_count >= 3:
                partial = round(0.7 * hidden_count / EXPECTED_HIDDEN_COUNT, 2)
                print(f"PARTIAL: Component 1 — {hidden_count}/{EXPECTED_HIDDEN_COUNT} expected rows hidden ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Only {hidden_count}/{EXPECTED_HIDDEN_COUNT} expected rows hidden (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: A count of hidden rows (5) is noted somewhere in the sheet (0.3 points)
    #
    # The task says "verify your work by checking the row count of visible rows"
    # and the context says "the total count of hidden rows is noted somewhere in the sheet."
    # In the golden_env, cell (21,1) contains 'Hidden rows count: 5'.
    # We look for a cell whose text content includes both the word "hidden" AND "5",
    # OR whose text explicitly states the count of hidden rows.
    # A plain numeric value of 5 in the data area (e.g., Score=5) must NOT trigger this.
    # -----------------------------------------------------------------------
    try:
        count_found = False
        count_details = None

        # Search all cells for a hidden row count annotation
        for row_num in range(1, ws.max_row + 1):
            for col_num in range(1, ws.max_column + 1):
                cell_val = ws.cell(row=row_num, column=col_num).value
                if cell_val is not None:
                    cell_str = str(cell_val).strip().lower()
                    # Must be a text cell (str type) containing "hidden" keyword + "5"
                    # This excludes plain numeric scores in the data area
                    if isinstance(cell_val, str) and 'hidden' in cell_str and '5' in cell_str:
                        count_found = True
                        count_details = f"Cell ({row_num},{col_num}): {cell_val!r}"
                        break
            if count_found:
                break

        if count_found:
            print(f"PASS: Component 2 — Hidden row count noted in sheet: {count_details} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — No hidden row count annotation found in sheet "
                  f"(expected text like 'Hidden rows count: 5' in a cell)")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
