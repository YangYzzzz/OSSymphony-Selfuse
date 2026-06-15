"""
Initial Setup: Survey Results spreadsheet with N/A and blank Score values
Task ID: osworld_calc_hide_rows_na_005
Domain: libreoffice_calc

Creates a survey results spreadsheet where some rows have N/A or blank Score values.
All rows are visible in the initial state (no rows hidden).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_hide_rows_na_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Survey Results ---
    ws = wb.active
    ws.title = "Survey Results"

    # Headers
    headers = ['Respondent ID', 'Name', 'Score', 'Feedback', 'Date']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, name='Calibri', size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 14

    # Survey data rows — mix of valid scores, N/A, and blank scores
    # Rows with N/A or blank: rows 3 (R003), 6 (R006), 9 (R009), 13 (R013), 16 (R016)
    # That's 5 rows to be hidden (will be captured in the golden patch)
    data = [
        # (Respondent ID, Name, Score, Feedback, Date)
        ('R001', 'Sarah Chen',        8,    'Very satisfied with the product features.',       '2025-01-10'),
        ('R002', 'Marcus Johnson',    7,    'Good overall, but support response could improve.','2025-01-11'),
        ('R003', 'Priya Patel',       'N/A','Did not complete the survey.',                    '2025-01-11'),
        ('R004', 'James Thornton',    9,    'Excellent experience from start to finish.',       '2025-01-12'),
        ('R005', 'Elena Rodriguez',   6,    'Average performance, some bugs encountered.',      '2025-01-13'),
        ('R006', 'David Kim',         '',   '',                                                 '2025-01-14'),
        ('R007', 'Natasha Williams',  10,   'Absolutely loved it! Highly recommended.',         '2025-01-14'),
        ('R008', 'Omar Hassan',       5,    'Mediocre. Expected better documentation.',         '2025-01-15'),
        ('R009', 'Linda Fischer',     'N/A','Survey link expired before completion.',           '2025-01-16'),
        ('R010', 'Brian Tanaka',      8,    'Solid product. Minor UI improvements needed.',     '2025-01-17'),
        ('R011', 'Angela Mwangi',     9,    'Great support team and fast onboarding.',          '2025-01-18'),
        ('R012', 'Steven Park',       7,    'Mostly positive experience with a few hiccups.',   '2025-01-19'),
        ('R013', 'Fatima Al-Rashid',  '',   '',                                                 '2025-01-19'),
        ('R014', 'Carlos Mendoza',    6,    'Needs improvement in the mobile app.',             '2025-01-20'),
        ('R015', 'Yuki Tanaka',       9,    'Intuitive interface and responsive service.',      '2025-01-21'),
        ('R016', 'Rachel Thompson',   'N/A','Technical difficulties prevented completion.',     '2025-01-22'),
        ('R017', 'Derek Sullivan',    8,    'Pleasant experience, would use again.',            '2025-01-23'),
        ('R018', 'Mei-Ling Zhou',     10,   'Exceeded all expectations. Outstanding product.',  '2025-01-24'),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # All rows are visible in initial state — NO rows hidden
    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
