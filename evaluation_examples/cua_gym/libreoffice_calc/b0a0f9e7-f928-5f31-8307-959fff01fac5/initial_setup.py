"""
Initial Setup: Create employees spreadsheet with 200 rows for AutoFilter task
Task ID: calc_gg5_007
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# --- Realistic data pools ---
FIRST_NAMES = [
    "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei", "Carlos",
    "Aisha", "Ryan", "Yuki", "Thomas", "Fatima", "Kevin", "Olga", "Nathan",
    "Zara", "Patrick", "Lin", "Ahmed", "Rachel", "Dmitri", "Sonia", "Brian",
    "Hana", "Trevor", "Amara", "Scott", "Nadia", "Vincent", "Kayla", "Miguel",
    "Ingrid", "Jason", "Deepa", "Robert", "Clara", "Omar", "Jessica", "Andre",
    "Mila", "Gregory", "Lena", "Derek", "Ava", "Samuel", "Tanya", "Luke",
    "Nina", "Felix"
]

LAST_NAMES = [
    "Chen", "Johnson", "Petrov", "Williams", "Sharma", "Kim", "Garcia",
    "Okafor", "Martinez", "Tanaka", "Singh", "Mueller", "Ali", "Brown",
    "Johansson", "Patel", "Lopez", "Yamamoto", "Nguyen", "Davis",
    "Anderson", "Costa", "Thompson", "Park", "Wilson", "Khan", "Lee",
    "Santos", "Robinson", "Eriksson", "Wright", "Fernandez", "Miller",
    "Sato", "Taylor", "Clark", "Wang", "Hall", "Morales", "Scott"
]

TITLES = [
    "Software Engineer", "Senior Software Engineer", "Staff Engineer",
    "Engineering Manager", "Product Manager", "Senior Product Manager",
    "Data Analyst", "Senior Data Analyst", "Data Scientist",
    "UX Designer", "Senior UX Designer", "Marketing Specialist",
    "Marketing Manager", "Sales Representative", "Account Executive",
    "HR Coordinator", "HR Manager", "Financial Analyst", "Senior Financial Analyst",
    "Operations Analyst", "Operations Manager", "Technical Writer",
    "QA Engineer", "Senior QA Engineer", "DevOps Engineer",
    "Security Engineer", "Business Analyst", "Project Manager",
    "Customer Success Manager", "Support Specialist"
]

LOCATIONS = [
    "San Francisco", "New York", "Austin", "Seattle", "Chicago",
    "Denver", "Boston", "Los Angeles", "Portland", "Atlanta",
    "London", "Toronto", "Singapore", "Berlin", "Tokyo"
]

DEPARTMENTS = [
    "Engineering", "Product", "Marketing", "Sales", "Human Resources",
    "Finance", "Operations", "Design", "Data Science", "Customer Support"
]

STATUSES = ["Active", "On Leave", "Contractor", "Probation"]


def create_initial():
    random.seed(42)  # reproducible data

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Directory"

    # Headers in row 1
    headers = ["ID", "Name", "Title", "Location", "Department", "Hire Date", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Generate 200 employee rows
    used_names = set()
    for i in range(1, 201):
        row = i + 1  # data starts at row 2

        # Employee ID
        emp_id = f"EMP-{1000 + i}"

        # Unique-ish name
        first = random.choice(FIRST_NAMES)
        last = random.choice(LAST_NAMES)
        name = f"{first} {last}"
        # Allow some duplicates in a large company, but try to avoid too many
        attempts = 0
        while name in used_names and attempts < 5:
            first = random.choice(FIRST_NAMES)
            last = random.choice(LAST_NAMES)
            name = f"{first} {last}"
            attempts += 1
        used_names.add(name)

        title = random.choice(TITLES)
        location = random.choice(LOCATIONS)
        department = random.choice(DEPARTMENTS)

        # Hire date between 2018-01-01 and 2025-12-31
        year = random.randint(2018, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        hire_date = f"{year}-{month:02d}-{day:02d}"

        # Status distribution: mostly Active
        status_weights = [0.70, 0.12, 0.10, 0.08]
        status = random.choices(STATUSES, weights=status_weights, k=1)[0]

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=title)
        ws.cell(row=row, column=4, value=location)
        ws.cell(row=row, column=5, value=department)
        ws.cell(row=row, column=6, value=hire_date)
        ws.cell(row=row, column=7, value=status)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14

    # NO AutoFilter -- that's the task
    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
