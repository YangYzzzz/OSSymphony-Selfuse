"""
Reward Script: AutoFilter on employees.xlsx — filter Engineering + Active
Task ID: calc_gg5_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): AutoFilter range is set on A1:G201
  Component 2 (0.25): FilterColumn on col E (colId=4) filters to 'Engineering'
  Component 3 (0.25): FilterColumn on col G (colId=6) filters to 'Active'
  Component 4 (0.25): Non-matching rows are hidden, matching rows are visible
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_007'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that AutoFilter is enabled on the Directory sheet with
    Department='Engineering' and Status='Active' filters applied.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Directory sheet
    if 'Directory' not in wb.sheetnames:
        print("CRITICAL: 'Directory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Directory']

    # Component 1: AutoFilter range is set (0.25 points)
    # The task asks to enable AutoFilter on A1:G200 (or A1:G201 including header).
    # We accept any autofilter ref that covers columns A through G and starts at row 1.
    try:
        af_ref = ws.auto_filter.ref
        if af_ref:
            # Parse the ref to check it covers the right columns
            ref_str = str(af_ref).upper()
            # Accept A1:G200, A1:G201, or similar that covers A-G from row 1
            if ref_str.startswith('A1:G') and ref_str.split(':')[1][1:].isdigit():
                end_row = int(ref_str.split(':')[1][1:])
                if end_row >= 200:
                    print(f"PASS: Component 1 — AutoFilter set on {af_ref} (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"FAIL: Component 1 — AutoFilter range ends at row {end_row}, expected >=200")
            else:
                print(f"FAIL: Component 1 — AutoFilter ref is '{af_ref}', expected A1:G200+")
        else:
            print("FAIL: Component 1 — AutoFilter not enabled (ref is None)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: FilterColumn on Department (colId=4) with 'Engineering' (0.25 points)
    try:
        dept_filter_vals = []
        for fc in ws.auto_filter.filterColumn:
            if fc.colId == 4:  # Column E = index 4 (0-based)
                if fc.filters:
                    dept_filter_vals = list(fc.filters.filter)
                break
        if 'Engineering' in dept_filter_vals:
            print(f"PASS: Component 2 — Department filter set to 'Engineering' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Expected 'Engineering' in colId=4 filters, found: {dept_filter_vals}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: FilterColumn on Status (colId=6) with 'Active' (0.25 points)
    try:
        status_filter_vals = []
        for fc in ws.auto_filter.filterColumn:
            if fc.colId == 6:  # Column G = index 6 (0-based)
                if fc.filters:
                    status_filter_vals = list(fc.filters.filter)
                break
        if 'Active' in status_filter_vals:
            print(f"PASS: Component 3 — Status filter set to 'Active' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Expected 'Active' in colId=6 filters, found: {status_filter_vals}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row visibility — non-matching rows hidden, matching rows visible (0.25 points)
    # Verify that rows with Department!='Engineering' or Status!='Active' are hidden,
    # and rows matching both criteria are visible.
    try:
        mismatches = 0
        total_data_rows = 0
        for r in range(2, ws.max_row + 1):
            dept = ws.cell(row=r, column=5).value
            status = ws.cell(row=r, column=7).value
            if dept is None and status is None:
                continue  # skip truly empty rows
            total_data_rows += 1
            is_hidden = ws.row_dimensions[r].hidden
            should_be_visible = (dept == 'Engineering' and status == 'Active')

            if should_be_visible and is_hidden:
                mismatches += 1  # matching row wrongly hidden
            elif not should_be_visible and not is_hidden:
                mismatches += 1  # non-matching row wrongly visible

        if total_data_rows == 0:
            print("FAIL: Component 4 — No data rows found")
        elif mismatches == 0:
            print(f"PASS: Component 4 — All {total_data_rows} rows correctly filtered (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — {mismatches} rows have incorrect visibility out of {total_data_rows}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
