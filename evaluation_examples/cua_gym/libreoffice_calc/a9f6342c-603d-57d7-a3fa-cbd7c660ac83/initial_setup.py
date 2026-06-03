"""
Initial Setup: Create Contact_Form spreadsheet with 39 contacts, no email validation
Task ID: calc_gcv_068
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contact_Form"

    # Headers
    ws.cell(row=1, column=1, value="Contact Name")
    ws.cell(row=1, column=2, value="Email")

    # Style headers
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col in [1, 2]:
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # 39 realistic contact names (A2:A40)
    contacts = [
        "Sarah Chen",
        "Marcus Johnson",
        "Priya Patel",
        "James O'Brien",
        "Fatima Al-Rashid",
        "Carlos Mendoza",
        "Emily Watson",
        "Takeshi Yamamoto",
        "Olivia Foster",
        "Raj Krishnamurthy",
        "Anna Kowalski",
        "David Thompson",
        "Sofia Hernandez",
        "Michael Chang",
        "Amara Okafor",
        "Lucas Bergstrom",
        "Isabella Rossi",
        "Nathan Goldberg",
        "Mei-Lin Wu",
        "Alexandra Petrov",
        "Hassan Demir",
        "Claire Dubois",
        "Benjamin Park",
        "Zara Mahmoud",
        "Ryan O'Sullivan",
        "Yuki Tanaka",
        "Grace Nkomo",
        "Daniel Fischer",
        "Leila Hosseini",
        "Thomas Andersen",
        "Naomi Sato",
        "Patrick Murphy",
        "Aisha Bah",
        "Viktor Novak",
        "Rachel Kim",
        "Eduardo Santos",
        "Helena Johansson",
        "Kevin Nguyen",
        "Simone Moreau",
    ]

    for i, name in enumerate(contacts, 2):
        ws.cell(row=i, column=1, value=name)
        # Column B (Email) left empty intentionally - no validation

    # Set column widths for readability
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 35

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
