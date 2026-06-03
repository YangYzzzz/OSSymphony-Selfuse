"""
Initial Setup: Budget template with categories and budget amounts
Task ID: calc_wf_002
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget"

    # --- Headers ---
    headers = ['Category', 'Budget', 'Actual', 'Variance']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # --- Data rows ---
    categories = [
        ('Rent', 1500),
        ('Utilities', 200),
        ('Groceries', 400),
        ('Transport', 150),
        ('Entertainment', 100),
        ('Savings', 500),
    ]

    for r, (cat, budget) in enumerate(categories, 2):
        # Column A: Category
        cell_a = ws.cell(row=r, column=1, value=cat)
        cell_a.font = Font(name="Calibri", size=11)
        cell_a.border = thin_border

        # Column B: Budget amount
        cell_b = ws.cell(row=r, column=2, value=budget)
        cell_b.font = Font(name="Calibri", size=11)
        cell_b.number_format = '$#,##0.00'
        cell_b.border = thin_border

        # Column C: Actual (empty - for user input)
        cell_c = ws.cell(row=r, column=3)
        cell_c.number_format = '$#,##0.00'
        cell_c.border = thin_border

        # Column D: Variance (empty in initial - task requires adding formula)
        cell_d = ws.cell(row=r, column=4)
        cell_d.number_format = '$#,##0.00'
        cell_d.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
