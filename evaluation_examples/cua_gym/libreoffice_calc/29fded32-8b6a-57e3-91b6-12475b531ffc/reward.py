"""
Reward Script: Budget template with variance formulas, conditional formatting, and sheet protection
Task ID: calc_wf_002
Domain: libreoffice_calc
Scoring:
  Component 1: Variance formulas in D2:D7 (0.35 pts)
  Component 2: Conditional formatting on D2:D7 — red for negative, green for positive (0.30 pts)
  Component 3: Sheet protection enabled (0.15 pts)
  Component 4: Only C2:C7 unlocked for editing (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_002'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Variance formulas in D2:D7 (0.35 points)
    # Task requires =B<n>-C<n> or =C<n>-B<n> in each D cell
    # Initial state: D2:D7 are all None — this check only passes on golden
    try:
        formula_count = 0
        for row in range(2, 8):
            cell_val = ws.cell(row=row, column=4).value  # column D
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                expected_a = f"=B{row}-C{row}"
                expected_b = f"=C{row}-B{row}"
                if normalized == expected_a or normalized == expected_b:
                    formula_count += 1
                else:
                    print(f"FAIL: D{row} has unexpected formula: {cell_val}")
            else:
                print(f"FAIL: D{row} has no formula (value: {cell_val!r})")

        if formula_count == 6:
            print(f"PASS: Component 1 — All 6 variance formulas correct (0.35 pts)")
            total_score += 0.35
        elif formula_count > 0:
            partial = round(0.35 * (formula_count / 6), 2)
            print(f"PARTIAL: Component 1 — {formula_count}/6 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No variance formulas found in D2:D7")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Conditional formatting on D2:D7 (0.30 points)
    # Requires two rules: red for over-budget (negative variance), green for under-budget (positive)
    # Initial state: no conditional formatting at all
    try:
        has_red_rule = False
        has_green_rule = False

        for cf in ws.conditional_formatting:
            # Check if the range covers D2:D7
            range_str = str(cf)
            # The range should include D column rows 2-7
            covers_d = False
            for cell_range in cf.cells.ranges:
                if (cell_range.min_col <= 4 <= cell_range.max_col and
                    cell_range.min_row <= 2 and cell_range.max_row >= 7):
                    covers_d = True
                    break

            if not covers_d:
                continue

            for rule in cf.rules:
                # Check for red rule (negative/less than 0)
                if rule.type == 'cellIs' and rule.operator in ('lessThan', 'lessThanOrEqual'):
                    formula_val = rule.formula[0] if rule.formula else ''
                    if '0' in str(formula_val):
                        has_red_rule = True
                        print(f"  Found red/negative rule: operator={rule.operator}, formula={rule.formula}")

                # Check for green rule (positive/greater than 0)
                if rule.type == 'cellIs' and rule.operator in ('greaterThan', 'greaterThanOrEqual'):
                    formula_val = rule.formula[0] if rule.formula else ''
                    if '0' in str(formula_val):
                        has_green_rule = True
                        print(f"  Found green/positive rule: operator={rule.operator}, formula={rule.formula}")

        if has_red_rule and has_green_rule:
            print(f"PASS: Component 2 — Both conditional formatting rules present (0.30 pts)")
            total_score += 0.30
        elif has_red_rule or has_green_rule:
            print(f"PARTIAL: Component 2 — Only one conditional formatting rule found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — No conditional formatting rules found on D2:D7")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sheet protection enabled (0.15 points)
    # Initial state: sheet is NOT protected
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 3 — Sheet protection is enabled (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Sheet protection is not enabled")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Only C2:C7 unlocked for editing (0.20 points)
    # Initial state: all cells are locked (default) and sheet is not protected
    # Golden state: C2:C7 are unlocked, all others remain locked, AND sheet is protected
    # We only score this if sheet protection is enabled (otherwise unlocking is meaningless)
    try:
        if not ws.protection.sheet:
            print(f"FAIL: Component 4 — Sheet not protected, so cell unlock status is irrelevant")
        else:
            c_unlocked_count = 0
            other_wrongly_unlocked = []

            # Check C2:C7 are unlocked
            for row in range(2, 8):
                cell = ws.cell(row=row, column=3)  # column C
                if not cell.protection.locked:
                    c_unlocked_count += 1

            # Spot-check that other important cells remain locked
            # Check A, B, D columns in rows 2-7 and headers
            for row in range(1, 8):
                for col in [1, 2, 4]:  # A, B, D
                    cell = ws.cell(row=row, column=col)
                    if not cell.protection.locked:
                        other_wrongly_unlocked.append(cell.coordinate)

            if c_unlocked_count == 6 and len(other_wrongly_unlocked) == 0:
                print(f"PASS: Component 4 — C2:C7 unlocked, all others locked (0.20 pts)")
                total_score += 0.20
            elif c_unlocked_count == 6:
                print(f"PARTIAL: Component 4 — C2:C7 unlocked but some others also unlocked: {other_wrongly_unlocked} (0.10 pts)")
                total_score += 0.10
            elif c_unlocked_count > 0:
                partial = round(0.20 * (c_unlocked_count / 6), 2)
                print(f"PARTIAL: Component 4 — {c_unlocked_count}/6 C cells unlocked ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — No C cells are unlocked")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
