"""
Initial Setup: Monthly Subscription and Recurring Expense Tracker
Task ID: calc_grs_090
Domain: libreoffice_calc

Creates a spreadsheet with 22 subscriptions/recurring expenses.
Raw data only - no formulas, no conditional formatting, no charts, no summaries.
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Subscriptions"

    # --- Headers ---
    headers = [
        "Service Name", "Category", "Monthly Cost", "Annual Cost",
        "Billing Cycle", "Next Billing Date", "Payment Method",
        "Notes", "Active Status", "Last Used"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Subscription Data (22 entries, sorted by Monthly Cost descending) ---
    # Columns: Service Name, Category, Monthly Cost, Annual Cost (blank - task asks for formula),
    #          Billing Cycle, Next Billing Date, Payment Method, Notes, Active (Y/N), Last Used
    today = date(2026, 4, 2)
    subscriptions = [
        ["Adobe Creative Cloud", "Software", 89.99, None, "Monthly", date(2026, 4, 8), "Visa ending 4521", "Full suite - design team use", "Y", date(2026, 4, 1)],
        ["Salesforce CRM", "Business", 75.00, None, "Monthly", date(2026, 4, 15), "Corporate Amex", "Enterprise license", "Y", date(2026, 4, 2)],
        ["Microsoft 365 Business", "Software", 59.99, None, "Monthly", date(2026, 4, 22), "Corporate Amex", "5-user plan", "Y", date(2026, 4, 2)],
        ["HubSpot Marketing Hub", "Business", 55.00, None, "Monthly", date(2026, 5, 1), "Visa ending 4521", "Professional tier", "Y", date(2026, 3, 28)],
        ["Peloton All-Access", "Health", 44.00, None, "Monthly", date(2026, 4, 5), "Mastercard ending 8832", "Family membership", "Y", date(2026, 3, 15)],
        ["LinkedIn Premium", "Business", 39.99, None, "Annual", date(2026, 8, 14), "Visa ending 4521", "Career plan", "Y", date(2026, 3, 20)],
        ["Spotify Family", "Entertainment", 16.99, None, "Monthly", date(2026, 4, 10), "PayPal", "6-person plan", "Y", date(2026, 4, 2)],
        ["Netflix Premium", "Entertainment", 15.49, None, "Monthly", date(2026, 4, 18), "Visa ending 4521", "4K streaming", "Y", date(2026, 4, 1)],
        ["iCloud+ 2TB", "Software", 14.99, None, "Monthly", date(2026, 4, 3), "Apple Pay", "Family sharing enabled", "Y", date(2026, 4, 2)],
        ["NordVPN", "Software", 12.99, None, "Annual", date(2026, 11, 5), "PayPal", "2-year plan billed annually", "Y", date(2026, 3, 30)],
        ["YouTube Premium", "Entertainment", 11.99, None, "Monthly", date(2026, 4, 12), "Visa ending 4521", "Family plan", "Y", date(2026, 4, 1)],
        ["Disney+ Bundle", "Entertainment", 10.99, None, "Monthly", date(2026, 4, 25), "Mastercard ending 8832", "Includes Hulu and ESPN+", "Y", date(2026, 3, 22)],
        ["Notion Plus", "Software", 10.00, None, "Monthly", date(2026, 4, 7), "Corporate Amex", "Team workspace", "Y", date(2026, 4, 2)],
        ["Amazon Prime", "Entertainment", 9.99, None, "Annual", date(2026, 7, 20), "Visa ending 4521", "Includes video and shipping", "Y", date(2026, 4, 2)],
        ["Calm Premium", "Health", 9.99, None, "Annual", date(2027, 1, 10), "PayPal", "Meditation and sleep stories", "Y", date(2026, 2, 14)],
        ["Home Chef Meal Kit", "Home", 9.99, None, "Monthly", date(2026, 4, 4), "Mastercard ending 8832", "Paused for summer travel", "N", date(2026, 2, 1)],
        ["New York Times Digital", "Entertainment", 9.99, None, "Monthly", date(2026, 4, 16), "Visa ending 4521", "All-access digital", "Y", date(2026, 4, 2)],
        ["Dropbox Plus", "Software", 9.99, None, "Annual", date(2026, 9, 1), "PayPal", "2TB storage", "Y", date(2026, 1, 20)],
        ["FitBod Gym App", "Health", 9.99, None, "Monthly", date(2026, 4, 9), "Apple Pay", "Signed up during pandemic - unused", "N", date(2024, 3, 15)],
        ["Grammarly Premium", "Software", 8.33, None, "Quarterly", date(2026, 5, 15), "Visa ending 4521", "Billed quarterly at $25", "Y", date(2026, 3, 25)],
        ["Ring Protect Plus", "Home", 8.33, None, "Annual", date(2026, 6, 18), "Mastercard ending 8832", "Covers all home cameras", "Y", date(2026, 4, 2)],
        ["Todoist Pro", "Software", 5.00, None, "Annual", date(2026, 12, 1), "PayPal", "Task management", "N", date(2025, 11, 10)],
    ]

    for r, row_data in enumerate(subscriptions, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:  # Monthly Cost
                cell.number_format = '$#,##0.00'
            elif c == 4:  # Annual Cost - leave blank
                pass
            elif c in (6, 10):  # Dates
                cell.number_format = 'yyyy-mm-dd'

    # --- Data Validations (Dropdowns) ---
    cat_dv = DataValidation(
        type="list",
        formula1='"Software,Entertainment,Health,Finance,Business,Home"',
        allow_blank=True,
        showDropDown=False,
    )
    cat_dv.prompt = "Select category"
    cat_dv.promptTitle = "Category"
    cat_dv.add("B2:B100")
    ws.add_data_validation(cat_dv)

    cycle_dv = DataValidation(
        type="list",
        formula1='"Monthly,Quarterly,Annual"',
        allow_blank=True,
        showDropDown=False,
    )
    cycle_dv.prompt = "Select billing cycle"
    cycle_dv.promptTitle = "Billing Cycle"
    cycle_dv.add("E2:E100")
    ws.add_data_validation(cycle_dv)

    # --- Column Widths ---
    col_widths = {
        "A": 28, "B": 16, "C": 14, "D": 14, "E": 14,
        "F": 18, "G": 24, "H": 38, "I": 14, "J": 14
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
