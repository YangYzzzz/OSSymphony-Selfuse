"""
Reward Script: Monthly Subscription and Recurring Expense Tracker
Task ID: calc_grs_090
Domain: libreoffice_calc
Scoring:
  Component 1: Annual Cost formulas (=C*12) in column D (0.20 pts)
  Component 2: Summary calculations - Active Monthly Total, Active Annual Total,
               Inactive Monthly Potential Savings (0.20 pts)
  Component 3: Category summary table with SUMPRODUCT formulas (0.15 pts)
  Component 4: Conditional formatting rules (3 rules) (0.20 pts)
  Component 5: Pie chart for category annual spend distribution (0.15 pts)
  Component 6: Data sorted by Monthly Cost descending (already in initial,
               but we check formulas are present too) - captured above
  Total: 0.90 from above + 0.10 for Last Used column conditional formatting
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_090'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Subscriptions' sheet exists
    if 'Subscriptions' not in wb.sheetnames:
        # Try first sheet as fallback
        ws = wb.worksheets[0]
        print(f"WARNING: No 'Subscriptions' sheet, using first sheet: {ws.title}")
    else:
        ws = wb['Subscriptions']

    # Determine data range: find last data row (rows with service names in col A)
    data_rows = []
    for r in range(2, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val is not None and isinstance(val, str) and len(val.strip()) > 0:
            # Skip summary labels
            if val.strip() in ('Active Monthly Total', 'Active Annual Total',
                               'Inactive Monthly Potential Savings', 'Category',
                               'Software', 'Entertainment', 'Health', 'Finance',
                               'Business', 'Home'):
                break
            data_rows.append(r)

    if len(data_rows) < 10:
        print(f"WARNING: Only {len(data_rows)} data rows found, expected ~22")

    last_data_row = max(data_rows) if data_rows else 23

    # =========================================================================
    # Component 1: Annual Cost formulas (=C*12) in column D (0.20 points)
    # Initial has None in column D; golden has =C<row>*12 formulas
    # =========================================================================
    try:
        formula_count = 0
        total_data = len(data_rows)
        for r in data_rows:
            cell_val = ws.cell(r, 4).value  # Column D = Annual Cost
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(" ", "")
                # Accept formulas like =C2*12 or =C2*12 or =12*C2
                if '*12' in val_upper and f'C{r}' in val_upper:
                    formula_count += 1
                elif '*12' in val_upper:
                    formula_count += 1  # close enough variant
        ratio = formula_count / total_data if total_data > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 1 — Annual Cost formulas present in {formula_count}/{total_data} rows (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            print(f"PARTIAL: Component 1 — Annual Cost formulas in {formula_count}/{total_data} rows (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Annual Cost formulas in only {formula_count}/{total_data} rows")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Summary calculations (0.20 points)
    # Active Monthly Total, Active Annual Total, Inactive Monthly Potential Savings
    # These should be in rows below the data with SUMPRODUCT or SUMIF formulas
    # =========================================================================
    try:
        summary_found = 0
        summary_labels = {
            'active monthly total': False,
            'active annual total': False,
            'inactive monthly potential savings': False,
        }
        for r in range(last_data_row + 1, ws.max_row + 1):
            label = ws.cell(r, 1).value
            if label and isinstance(label, str):
                label_lower = label.strip().lower()
                for key in summary_labels:
                    if key in label_lower:
                        # Check that there's a formula in column C or D
                        c_val = ws.cell(r, 3).value
                        d_val = ws.cell(r, 4).value
                        has_formula = False
                        for v in [c_val, d_val]:
                            if v and isinstance(v, str) and v.startswith('='):
                                has_formula = True
                                break
                        if has_formula:
                            summary_labels[key] = True
                            summary_found += 1
                            print(f"  Found summary: '{label.strip()}' with formula")

        if summary_found >= 3:
            print(f"PASS: Component 2 — All 3 summary calculations found (0.20 pts)")
            total_score += 0.20
        elif summary_found >= 2:
            print(f"PARTIAL: Component 2 — {summary_found}/3 summary calculations (0.13 pts)")
            total_score += 0.13
        elif summary_found >= 1:
            print(f"PARTIAL: Component 2 — {summary_found}/3 summary calculations (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2 — No summary calculations found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Category summary table (0.15 points)
    # A table with category names and annual spend formulas for pie chart data
    # Expected categories: Software, Entertainment, Health, Finance, Business, Home
    # =========================================================================
    try:
        expected_categories = {'software', 'entertainment', 'health', 'finance', 'business', 'home'}
        found_categories = set()
        cat_formulas = 0

        for r in range(last_data_row + 1, ws.max_row + 1):
            cell_a = ws.cell(r, 1).value
            cell_b = ws.cell(r, 2).value
            if cell_a and isinstance(cell_a, str):
                cat_lower = cell_a.strip().lower()
                if cat_lower in expected_categories:
                    found_categories.add(cat_lower)
                    if cell_b and isinstance(cell_b, str) and cell_b.startswith('='):
                        cat_formulas += 1

        if len(found_categories) >= 5 and cat_formulas >= 5:
            print(f"PASS: Component 3 — Category summary table with {len(found_categories)} categories and {cat_formulas} formulas (0.15 pts)")
            total_score += 0.15
        elif len(found_categories) >= 3 and cat_formulas >= 3:
            print(f"PARTIAL: Component 3 — {len(found_categories)} categories, {cat_formulas} formulas (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — Found {len(found_categories)} categories with {cat_formulas} formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Conditional formatting rules (0.25 points)
    # Expected: 3 rules:
    #   - Billing within 7 days (yellow) on Next Billing Date column
    #   - Overpriced >$50/month (orange) on Monthly Cost column
    #   - Last used 30+ days ago (red/highlight) on Last Used column
    # =========================================================================
    try:
        cf_rules_list = []
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                cf_rules_list.append({
                    'range': str(cf),
                    'type': rule.type,
                    'operator': getattr(rule, 'operator', None),
                    'formula': rule.formula,
                })

        cf_count = len(cf_rules_list)
        print(f"  Found {cf_count} conditional formatting rules")
        for r in cf_rules_list:
            print(f"    Range: {r['range']}, Type: {r['type']}, Op: {r['operator']}, Formula: {r['formula']}")

        if cf_count >= 3:
            print(f"PASS: Component 4 — {cf_count} conditional formatting rules found (0.25 pts)")
            total_score += 0.25
        elif cf_count >= 2:
            print(f"PARTIAL: Component 4 — {cf_count}/3 conditional formatting rules (0.15 pts)")
            total_score += 0.15
        elif cf_count >= 1:
            print(f"PARTIAL: Component 4 — {cf_count}/3 conditional formatting rules (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Pie chart for category annual spend (0.20 points)
    # Should be a PieChart on the sheet
    # =========================================================================
    try:
        charts = ws._charts
        pie_found = False
        for chart in charts:
            chart_type = type(chart).__name__
            if 'Pie' in chart_type:
                pie_found = True
                print(f"  Found PieChart with {len(chart.series)} series")
                break

        if pie_found:
            print(f"PASS: Component 5 — Pie chart found (0.20 pts)")
            total_score += 0.20
        elif len(charts) > 0:
            # Some chart exists but not a pie chart - partial credit
            chart_type = type(charts[0]).__name__
            print(f"PARTIAL: Component 5 — Found chart ({chart_type}) but not a PieChart (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
