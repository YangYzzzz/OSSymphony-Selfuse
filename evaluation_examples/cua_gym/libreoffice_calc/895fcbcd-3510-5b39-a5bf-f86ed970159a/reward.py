"""
Reward Script: Insert bar-type sparklines in column H for 30 rows of monthly data
Task ID: calc_gcp_072
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): All 30 cells H2:H31 contain a formula
  Component 2 (0.3): All formulas are SPARKLINE formulas referencing correct row data range (B:G)
  Component 3 (0.3): All SPARKLINE formulas specify bar chart type
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_072'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'ProductMetrics' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ProductMetrics' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ProductMetrics']

    # Component 1: All 30 cells H2:H31 contain a formula (0.4 points)
    # This checks that the agent populated column H with formulas.
    # Initial env has None in all H cells, so this fails on initial.
    try:
        formula_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=8).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_count += 1
        ratio = formula_count / 30.0
        if ratio >= 1.0:
            print(f"PASS: Component 1 — All 30 cells H2:H31 contain formulas (0.4 pts)")
            total_score += 0.4
        elif ratio > 0:
            partial = round(0.4 * ratio, 2)
            print(f"PARTIAL: Component 1 — {formula_count}/30 cells have formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No formulas found in H2:H31")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All formulas are SPARKLINE formulas referencing the correct row range B:G (0.3 points)
    # Each cell Hn should have a SPARKLINE formula that references Bn:Gn (the 6-month data for that row).
    # Initial env has no formulas, so this fails on initial.
    try:
        sparkline_correct_range_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=8).value
            if val is not None and isinstance(val, str):
                upper_val = val.upper()
                # Check it's a SPARKLINE function
                if '=SPARKLINE(' not in upper_val:
                    continue
                # Check it references the correct row range B<r>:G<r>
                expected_range = f"B{r}:G{r}".upper()
                if expected_range in upper_val:
                    sparkline_correct_range_count += 1
        ratio = sparkline_correct_range_count / 30.0
        if ratio >= 1.0:
            print(f"PASS: Component 2 — All 30 SPARKLINE formulas reference correct row ranges (0.3 pts)")
            total_score += 0.3
        elif ratio > 0:
            partial = round(0.3 * ratio, 2)
            print(f"PARTIAL: Component 2 — {sparkline_correct_range_count}/30 have correct SPARKLINE ranges ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No correct SPARKLINE range references found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All SPARKLINE formulas use bar chart type (0.3 points)
    # The task specifically requires bar-type sparklines (not line-type).
    # We check for "bar" in the charttype parameter of the SPARKLINE function.
    # Initial env has no formulas, so this fails on initial.
    try:
        bar_type_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=8).value
            if val is not None and isinstance(val, str):
                upper_val = val.upper()
                if '=SPARKLINE(' not in upper_val:
                    continue
                # Check for bar chart type specification
                # Acceptable patterns: "charttype","bar" or "charttype";"bar" or charttype=bar etc.
                # LibreOffice SPARKLINE uses: =SPARKLINE(range,{"charttype","bar"})
                lower_val = val.lower()
                if '"bar"' in lower_val or "'bar'" in lower_val or ',bar' in lower_val.replace(' ', ''):
                    bar_type_count += 1
        ratio = bar_type_count / 30.0
        if ratio >= 1.0:
            print(f"PASS: Component 3 — All 30 sparklines are bar-type (0.3 pts)")
            total_score += 0.3
        elif ratio > 0:
            partial = round(0.3 * ratio, 2)
            print(f"PARTIAL: Component 3 — {bar_type_count}/30 sparklines are bar-type ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No bar-type sparklines found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
