"""
Initial Setup: Create ProductMetrics spreadsheet with 30 products and 6 months of sales data.
Task ID: calc_gcp_072
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_072'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProductMetrics'

    # --- Headers ---
    headers = ['Product', 'Month1', 'Month2', 'Month3', 'Month4', 'Month5', 'Month6', 'Sparkline']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Product names (30 realistic products) ---
    products = [
        'Alpine Trail Boots', 'Cascade Water Filter', 'Delta Running Shoes',
        'Echo Bluetooth Speaker', 'Falcon Drone Pro', 'Granite Cookware Set',
        'Harbor LED Lantern', 'Ironclad Tool Belt', 'Jade Yoga Mat',
        'Keystone Binoculars', 'Lakeview Tent 4P', 'Meridian Sunglasses',
        'Nimbus Rain Jacket', 'Orbit Fitness Tracker', 'Pioneer Backpack 40L',
        'Quantum Thermos Flask', 'Ridge Hiking Poles', 'Summit Sleeping Bag',
        'Tidal Kayak Paddle', 'Ultra Headlamp 600', 'Venture Compass Kit',
        'Wildfire Camp Stove', 'Xenon Power Bank', 'Yukon Cooler Box',
        'Zenith Climbing Rope', 'Aurora Down Jacket', 'Blaze Fire Starter',
        'Crest Mountain Bike', 'Drift Snorkel Set', 'Eclipse Solar Charger',
    ]

    # --- Generate 6 months of sales data with varied patterns ---
    # Some products show growth, some decline, some stable, some volatile
    patterns = ['growth', 'decline', 'stable', 'volatile', 'seasonal']

    data_font = Font(name='Calibri', size=11)
    product_align = Alignment(horizontal='left', vertical='center')
    number_align = Alignment(horizontal='center', vertical='center')

    for i, product in enumerate(products):
        row = i + 2
        pattern = patterns[i % len(patterns)]
        base = random.randint(80, 350)

        monthly_values = []
        for m in range(6):
            if pattern == 'growth':
                val = base + m * random.randint(15, 40) + random.randint(-10, 10)
            elif pattern == 'decline':
                val = base - m * random.randint(15, 35) + random.randint(-10, 10)
            elif pattern == 'stable':
                val = base + random.randint(-20, 20)
            elif pattern == 'volatile':
                val = random.randint(50, 500)
            else:  # seasonal
                seasonal_bump = [0.7, 0.8, 1.0, 1.3, 1.5, 1.2]
                val = int(base * seasonal_bump[m] + random.randint(-15, 15))
            val = max(50, min(500, val))
            monthly_values.append(val)

        # Write product name
        cell_a = ws.cell(row=row, column=1, value=product)
        cell_a.font = data_font
        cell_a.alignment = product_align
        cell_a.border = thin_border

        # Write monthly values
        for c, val in enumerate(monthly_values, 2):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = data_font
            cell.alignment = number_align
            cell.number_format = '#,##0'
            cell.border = thin_border

        # Column H (Sparkline) - leave EMPTY in initial state
        cell_h = ws.cell(row=row, column=8)
        cell_h.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 24
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 10
    ws.column_dimensions['H'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
