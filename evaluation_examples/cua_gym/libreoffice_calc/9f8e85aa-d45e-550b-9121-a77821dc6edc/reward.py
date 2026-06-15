"""
Reward Script: Standardize region names in SalesDB and build Validation summary
Task ID: calc_gen_data_cleanup_071
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: All region values in SalesDB column D are standardized (0.5 pts)
  Component 2: Validation sheet exists with correct structure and COUNTIF formulas (0.3 pts)
  Component 3: SalesDB is sorted by Region then Date (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_071'

STANDARD_REGIONS = {'Northeast', 'Southeast', 'Midwest', 'West'}

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: SalesDB sheet must exist
    if 'SalesDB' not in wb.sheetnames:
        print("CRITICAL: Sheet 'SalesDB' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws_sales = wb['SalesDB']

    # Component 1: All region values in column D are standardized to one of the 4 standard names (0.5 pts)
    # Initial file has 20 distinct variant values. After task: only 4 standard values should remain.
    try:
        non_standard = []
        total_rows = 0
        for row in range(2, ws_sales.max_row + 1):
            val = ws_sales.cell(row=row, column=4).value
            if val is not None:
                total_rows += 1
                if val not in STANDARD_REGIONS:
                    non_standard.append((row, val))

        if total_rows == 0:
            print("FAIL: Component 1 — No data rows found in SalesDB column D")
        elif len(non_standard) == 0 and total_rows == 3000:
            print(f"PASS: Component 1 — All {total_rows} region values are standardized to one of {STANDARD_REGIONS} (0.5 pts)")
            total_score += 0.5
        elif len(non_standard) == 0:
            print(f"FAIL: Component 1 — Region values are standardized but row count is {total_rows}, expected 3000")
        else:
            print(f"FAIL: Component 1 — {len(non_standard)} non-standard region values found (e.g. row {non_standard[0][0]}: {repr(non_standard[0][1])})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Validation sheet exists with correct structure and COUNTIF formulas (0.3 pts)
    # Expects sheet 'Validation' with:
    #   A1: 'Region', B1: 'Count'
    #   A2-A5: 4 standard region names, B2-B5: COUNTIF formulas for SalesDB column D
    #   A7: 'Total', B7: SUM formula
    #   A8: 'Unmatched', B8: COUNTA-SUM formula
    try:
        if 'Validation' not in wb.sheetnames:
            print("FAIL: Component 2 — Sheet 'Validation' does not exist")
        else:
            ws_val = wb['Validation']
            validation_score = 0.0

            # Check headers
            hdr_a = ws_val['A1'].value
            hdr_b = ws_val['B1'].value
            has_headers = (
                hdr_a is not None and str(hdr_a).strip().lower() == 'region' and
                hdr_b is not None and str(hdr_b).strip().lower() == 'count'
            )

            # Check 4 region rows with COUNTIF formulas referencing SalesDB
            region_labels = set()
            countif_count = 0
            for row in range(2, 6):
                label = ws_val.cell(row=row, column=1).value
                formula = ws_val.cell(row=row, column=2).value
                if label is not None:
                    region_labels.add(str(label).strip())
                if (formula is not None and
                        isinstance(formula, str) and
                        formula.upper().startswith('=COUNTIF')):
                    countif_count += 1

            has_four_regions = region_labels == STANDARD_REGIONS
            has_four_countifs = (countif_count == 4)

            # Check Total row (row 7) and Unmatched row (row 8)
            total_label = ws_val.cell(row=7, column=1).value
            total_formula = ws_val.cell(row=7, column=2).value
            unmatched_label = ws_val.cell(row=8, column=1).value
            unmatched_formula = ws_val.cell(row=8, column=2).value

            has_total_row = (
                total_label is not None and str(total_label).strip().lower() == 'total' and
                total_formula is not None and isinstance(total_formula, str) and
                total_formula.upper().startswith('=SUM')
            )
            has_unmatched_row = (
                unmatched_label is not None and str(unmatched_label).strip().lower() == 'unmatched' and
                unmatched_formula is not None and isinstance(unmatched_formula, str) and
                '=' in unmatched_formula
            )

            passed_checks = []
            failed_checks = []

            if has_headers:
                passed_checks.append("headers(Region/Count)")
            else:
                failed_checks.append(f"headers (got A1={repr(hdr_a)}, B1={repr(hdr_b)})")

            if has_four_regions:
                passed_checks.append("4 region labels")
            else:
                failed_checks.append(f"4 region labels (found: {region_labels})")

            if has_four_countifs:
                passed_checks.append("4 COUNTIF formulas")
            else:
                failed_checks.append(f"4 COUNTIF formulas (found {countif_count})")

            if has_total_row:
                passed_checks.append("Total/SUM row")
            else:
                failed_checks.append(f"Total/SUM row (got label={repr(total_label)}, formula={repr(total_formula)})")

            if has_unmatched_row:
                passed_checks.append("Unmatched row")
            else:
                failed_checks.append(f"Unmatched row (got label={repr(unmatched_label)}, formula={repr(unmatched_formula)})")

            # Award 0.3 only if all structure checks pass
            if not failed_checks:
                print(f"PASS: Component 2 — Validation sheet has correct structure: {', '.join(passed_checks)} (0.3 pts)")
                total_score += 0.3
            else:
                # Partial: 0.15 if sheet exists and has the 4 region labels+countifs even if missing total/unmatched
                if has_four_regions and has_four_countifs:
                    print(f"PARTIAL: Component 2 — Validation sheet has regions+countifs but missing: {', '.join(failed_checks)} (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 2 — Validation sheet issues: {'; '.join(failed_checks)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: SalesDB is sorted by Region then Date (0.2 pts)
    # Initial file is NOT sorted. Golden file IS sorted by Region then Date.
    try:
        rows = []
        for row in range(2, ws_sales.max_row + 1):
            region = ws_sales.cell(row=row, column=4).value
            date = ws_sales.cell(row=row, column=2).value
            if region is not None:
                rows.append((str(region), date))

        if len(rows) < 2:
            print("FAIL: Component 3 — Not enough rows to check sort order")
        else:
            out_of_order = 0
            for i in range(len(rows) - 1):
                r1, d1 = rows[i]
                r2, d2 = rows[i + 1]
                # Region must be non-decreasing
                if r1 > r2:
                    out_of_order += 1
                elif r1 == r2:
                    # Within same region, date must be non-decreasing
                    try:
                        if d1 is not None and d2 is not None and d1 > d2:
                            out_of_order += 1
                    except TypeError:
                        pass  # Skip non-comparable dates

            if out_of_order == 0:
                print(f"PASS: Component 3 — SalesDB is sorted by Region then Date (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — SalesDB has {out_of_order} out-of-order row pair(s)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
