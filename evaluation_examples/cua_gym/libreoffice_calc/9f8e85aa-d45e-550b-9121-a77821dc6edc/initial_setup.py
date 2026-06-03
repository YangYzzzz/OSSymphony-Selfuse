"""
Initial Setup: Sales database with inconsistent region names for standardization task
Task ID: calc_gen_data_cleanup_071
Domain: libreoffice_calc
"""

import random
from datetime import date, timedelta
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Seed for reproducibility
random.seed(42)

# Region variant mappings (inconsistent names per region)
REGION_VARIANTS = {
    'Northeast': ['NE', 'N.E.', 'North East', 'NORTHEAST', 'Northeast', 'north east'],
    'Southeast': ['SE', 'S.E.', 'South East', 'SOUTHEAST', 'Southeast'],
    'Midwest':   ['MW', 'Mid-West', 'MIDWEST', 'Midwest', 'midwest'],
    'West':      ['W', 'WEST', 'West', 'west'],
}

# Flatten into a pool weighted roughly equally across 4 regions
region_pool = []
for region, variants in REGION_VARIANTS.items():
    for v in variants:
        region_pool.extend([v] * 4)  # repeat each variant a few times for plausibility

# Sales representatives
reps = [
    'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
    'Jessica Patel', 'Ryan Thompson', 'Amanda Brooks', 'Carlos Mendez',
    'Natalie Wong', 'Brian Foster', 'Lauren Mitchell', 'James Carter',
    'Stephanie Hayes', 'Kevin Liu', 'Rachel Adams',
]

# Products
products = [
    'Enterprise Suite', 'Pro License', 'Starter Pack', 'Support Contract',
    'Analytics Module', 'Cloud Storage', 'Security Bundle', 'API Access',
    'Training Package', 'Consulting Hours',
]

def random_date(start_year=2023, end_year=2025):
    start = date(start_year, 1, 1)
    end = date(end_year, 12, 31)
    delta = (end - start).days
    return start + timedelta(days=random.randint(0, delta))

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: SalesDB ---
    ws = wb.active
    ws.title = 'SalesDB'

    # Headers
    headers = ['Trans ID', 'Date', 'Rep', 'Region', 'Product', 'Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 3000 rows of transaction data
    num_rows = 3000
    for i in range(num_rows):
        trans_id = f'TXN-{10001 + i}'
        txn_date = random_date()
        rep = random.choice(reps)
        region = random.choice(region_pool)
        product = random.choice(products)
        amount = round(random.uniform(500, 75000), 2)

        row_num = i + 2
        ws.cell(row=row_num, column=1, value=trans_id)
        ws.cell(row=row_num, column=2, value=txn_date.strftime('%Y-%m-%d'))
        ws.cell(row=row_num, column=3, value=rep)
        ws.cell(row=row_num, column=4, value=region)
        ws.cell(row=row_num, column=5, value=product)
        ws.cell(row=row_num, column=6, value=amount)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  SalesDB rows: {num_rows} (plus 1 header = {num_rows + 1} total)')
    print(f'  Region variants used: {len(region_pool)} entries in pool')

create_initial()
