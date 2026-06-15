"""
Reward Script: Build two pivot tables in Summary sheet from clinical trial data
Task ID: osworld_calc_pivot_dual_dimensions_005
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.20): Summary sheet (Sheet2) exists and is non-empty with pivot table content
  Component 2 (0.40): First pivot table — Patient Count by Trial Site — present with correct values
  Component 3 (0.40): Second pivot table — Average Efficacy Score by Drug Dosage Level — correct values
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_dual_dimensions_005'

# Ground truth values derived from task context and source data
EXPECTED_SITE_COUNTS = {
    'Boston Medical Center': 5,
    'Chicago Health Network': 5,
    'Houston Research Institute': 5,
    'Miami Clinical Hub': 5,
    'Seattle Cancer Center': 5,
}
EXPECTED_TOTAL_PATIENTS = 25

# Exact average efficacy scores from golden artifact (verified via VM exploration)
EXPECTED_DOSAGE_EXACT = {
    'High': 89.03,
    'Low': 60.39,
    'Medium': 74.86,
}
EXPECTED_DOSAGE_TOTAL = 74.18
TOLERANCE = 0.5  # allow slight rounding differences


def find_sheet2(wb):
    """
    Find the summary/pivot sheet. Task refers to 'Sheet2', but the actual name may vary.
    Returns the worksheet or None.
    """
    if 'Sheet2' in wb.sheetnames:
        return wb['Sheet2']
    if 'Summary' in wb.sheetnames:
        return wb['Summary']
    # Return any non-data sheet
    data_sheets = {'Sheet1', 'ClinicalTrials'}
    for name in wb.sheetnames:
        if name not in data_sheets:
            return wb[name]
    return None


def scan_pivot_tables(ws):
    """
    Scan the Summary sheet and extract two dictionaries:
      - site_counts: {site_name: count}
      - dosage_avgs: {dosage_level: average_score}
    Also extract grand totals.
    Returns: (site_counts, site_total, dosage_avgs, dosage_total)

    Layout expected (based on golden artifact):
      Row 1:  Title: "Patient Count by Trial Site"
      Row 2:  Headers: "Trial Site" | "Patient Count"
      Rows 3-7: Data rows (site_name, count)
      Row 8:  ("Grand Total", 25)
      Row 9:  (empty)
      Row 10: Title: "Average Efficacy Score by Drug Dosage Level"
      Row 11: Headers: "Drug Dosage Level" | "Average Efficacy Score"
      Rows 12-14: Data rows (dosage_level, avg_score)
      Row 15: ("Grand Total", 74.18)
    """
    site_counts = {}
    site_total = None
    dosage_avgs = {}
    dosage_total = None

    # Use index-based state machine to avoid re-triggering on header column text
    # States: 'idle', 'site_data', 'dosage_data'
    state = 'idle'
    # Track the last seen title to know which section follows
    last_title = None

    for row in ws.iter_rows(values_only=True):
        row_text = [str(v).lower().strip() if v is not None else '' for v in row]
        non_none = [(i, v) for i, v in enumerate(row) if v is not None]

        if not non_none:
            # Empty rows separate the two pivot tables
            state = 'idle'
            continue

        # --- Title row detection (only check single-cell or first-cell content) ---
        # Title rows have content only in the first cell
        first_text = row_text[0] if row_text else ''

        if 'patient count by trial site' in first_text:
            last_title = 'site'
            state = 'idle'  # next non-empty row is header
            continue

        if 'average efficacy' in first_text and 'drug dosage' not in first_text:
            last_title = 'dosage'
            state = 'idle'
            continue

        # --- Column header row detection ---
        # Triggered when we just saw a title row (last_title is set)
        if last_title == 'site' and state == 'idle':
            if any('trial site' in t for t in row_text):
                state = 'site_data'
                last_title = None
                continue

        if last_title == 'dosage' and state == 'idle':
            if any('drug dosage' in t for t in row_text):
                state = 'dosage_data'
                last_title = None
                continue

        # --- Fallback: detect column headers directly if no title was seen ---
        if state == 'idle':
            if any('trial site' in t for t in row_text) and any('count' in t or 'patient' in t for t in row_text):
                state = 'site_data'
                continue
            if any('drug dosage' in t for t in row_text) and any('efficacy' in t or 'average' in t for t in row_text):
                state = 'dosage_data'
                continue

        # --- Data row parsing ---
        if state == 'site_data' and len(non_none) >= 2:
            first_val = non_none[0][1]
            second_val = non_none[1][1]
            if isinstance(first_val, str) and 'total' in first_val.lower():
                try:
                    site_total = int(second_val)
                except (TypeError, ValueError):
                    pass
            elif isinstance(first_val, str) and isinstance(second_val, (int, float)):
                site_counts[first_val] = int(second_val)

        elif state == 'dosage_data' and len(non_none) >= 2:
            first_val = non_none[0][1]
            second_val = non_none[1][1]
            if isinstance(first_val, str) and 'total' in first_val.lower():
                try:
                    dosage_total = float(second_val)
                except (TypeError, ValueError):
                    pass
            elif isinstance(first_val, str) and isinstance(second_val, (int, float)):
                dosage_avgs[first_val] = float(second_val)

    return site_counts, site_total, dosage_avgs, dosage_total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Component 1: Summary sheet exists and is non-empty with pivot content (0.20 pts)
    # -------------------------------------------------------------------------
    try:
        ws_summary = find_sheet2(wb)
        if ws_summary is None:
            print("FAIL: Component 1 — No summary/Sheet2 found in workbook")
            print(f"  Sheets present: {wb.sheetnames}")
        else:
            all_values = [(cell.row, cell.column, cell.value)
                          for row in ws_summary.iter_rows()
                          for cell in row if cell.value is not None]
            if len(all_values) >= 10:
                print(f"PASS: Component 1 — Summary sheet '{ws_summary.title}' found with {len(all_values)} non-empty cells (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — Summary sheet '{ws_summary.title}' has only {len(all_values)} non-empty cells (need >= 10)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        ws_summary = None

    if ws_summary is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {min(total_score, 1.0)}")
        return min(total_score, 1.0)

    # Parse pivot table data
    try:
        site_counts, site_total, dosage_avgs, dosage_total = scan_pivot_tables(ws_summary)
    except Exception as e:
        print(f"ERROR: Failed to parse pivot tables — {e}")
        site_counts, site_total, dosage_avgs, dosage_total = {}, None, {}, None

    # -------------------------------------------------------------------------
    # Component 2: Patient Count by Trial Site pivot table (0.40 pts)
    # -------------------------------------------------------------------------
    try:
        sites_correct = sum(
            1 for site, expected in EXPECTED_SITE_COUNTS.items()
            if site_counts.get(site) == expected
        )
        sites_expected = len(EXPECTED_SITE_COUNTS)
        total_ok = (site_total == EXPECTED_TOTAL_PATIENTS)

        # Log any mismatches
        for site, expected in EXPECTED_SITE_COUNTS.items():
            actual = site_counts.get(site, 'NOT FOUND')
            if actual != expected:
                print(f"  SITE CHECK: '{site}' expected={expected}, found={actual}")

        if sites_correct == sites_expected and total_ok:
            print(f"PASS: Component 2 — All {sites_expected} trial site counts correct + Grand Total={site_total} (0.40 pts)")
            total_score += 0.40
        elif sites_correct >= 3 and total_ok:
            print(f"PARTIAL: Component 2 — {sites_correct}/{sites_expected} site counts correct, Grand Total={site_total} (0.25 pts)")
            total_score += 0.25
        elif sites_correct >= 3:
            print(f"PARTIAL: Component 2 — {sites_correct}/{sites_expected} site counts correct, Grand Total incorrect/missing (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Only {sites_correct}/{sites_expected} site counts correct, Grand Total={site_total}")
            print(f"  Found site counts: {site_counts}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Average Efficacy Score by Drug Dosage Level pivot table (0.40 pts)
    # -------------------------------------------------------------------------
    try:
        dosage_correct = sum(
            1 for dosage, expected_avg in EXPECTED_DOSAGE_EXACT.items()
            if dosage in dosage_avgs and abs(dosage_avgs[dosage] - expected_avg) <= TOLERANCE
        )
        dosage_expected = len(EXPECTED_DOSAGE_EXACT)
        grand_total_ok = (
            dosage_total is not None and
            abs(dosage_total - EXPECTED_DOSAGE_TOTAL) <= TOLERANCE
        )

        # Log any mismatches
        for dosage, expected_avg in EXPECTED_DOSAGE_EXACT.items():
            actual = dosage_avgs.get(dosage, 'NOT FOUND')
            if actual == 'NOT FOUND':
                print(f"  DOSAGE CHECK: '{dosage}' NOT FOUND in dosage table")
            elif abs(float(actual) - expected_avg) > TOLERANCE:
                print(f"  DOSAGE CHECK: '{dosage}' expected~{expected_avg}, found={actual:.4f}")

        if dosage_correct == dosage_expected and grand_total_ok:
            print(f"PASS: Component 3 — All {dosage_expected} dosage averages correct + Grand Total={dosage_total:.2f} (0.40 pts)")
            total_score += 0.40
        elif dosage_correct == dosage_expected:
            print(f"PARTIAL: Component 3 — All {dosage_expected} dosage averages correct, Grand Total={dosage_total} (0.30 pts)")
            total_score += 0.30
        elif dosage_correct >= 2:
            print(f"PARTIAL: Component 3 — {dosage_correct}/{dosage_expected} dosage averages correct (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Only {dosage_correct}/{dosage_expected} dosage averages correct")
            print(f"  Found dosage avgs: {dosage_avgs}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path on VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
