"""
Initial Setup: Clinical trial data spreadsheet for pivot table creation task.
Task ID: osworld_calc_pivot_dual_dimensions_005
Domain: libreoffice_calc

Creates:
  - Sheet1 (ClinicalTrials): 25 rows of clinical trial records with columns
    Patient ID, Trial Site, Drug Dosage Level, Efficacy Score, Treatment Duration
  - Sheet2 (Summary): Empty sheet (agent must build pivot tables here)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_dual_dimensions_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: ClinicalTrials ---
    ws1 = wb.active
    ws1.title = 'ClinicalTrials'

    # Headers
    headers = ['Patient ID', 'Trial Site', 'Drug Dosage Level', 'Efficacy Score', 'Treatment Duration (days)']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Clinical trial data — realistic patient records
    # Sites: Boston Medical Center, Houston Research Institute, Seattle Cancer Center, Chicago Health Network, Miami Clinical Hub
    # Dosage levels: Low, Medium, High
    data = [
        ['PT-001', 'Boston Medical Center',        'Low',    62.3, 28],
        ['PT-002', 'Houston Research Institute',   'High',   88.7, 21],
        ['PT-003', 'Seattle Cancer Center',        'Medium', 74.1, 35],
        ['PT-004', 'Chicago Health Network',       'Low',    55.8, 42],
        ['PT-005', 'Miami Clinical Hub',           'High',   91.2, 14],
        ['PT-006', 'Boston Medical Center',        'Medium', 78.5, 30],
        ['PT-007', 'Houston Research Institute',   'Low',    61.0, 45],
        ['PT-008', 'Seattle Cancer Center',        'High',   85.4, 18],
        ['PT-009', 'Chicago Health Network',       'Medium', 72.6, 33],
        ['PT-010', 'Miami Clinical Hub',           'Low',    58.3, 40],
        ['PT-011', 'Boston Medical Center',        'High',   89.9, 17],
        ['PT-012', 'Houston Research Institute',   'Medium', 76.2, 29],
        ['PT-013', 'Seattle Cancer Center',        'Low',    63.7, 44],
        ['PT-014', 'Chicago Health Network',       'High',   87.1, 20],
        ['PT-015', 'Miami Clinical Hub',           'Medium', 71.8, 36],
        ['PT-016', 'Boston Medical Center',        'Low',    59.4, 38],
        ['PT-017', 'Houston Research Institute',   'High',   92.6, 12],
        ['PT-018', 'Seattle Cancer Center',        'Medium', 75.3, 31],
        ['PT-019', 'Chicago Health Network',       'Low',    60.9, 43],
        ['PT-020', 'Miami Clinical Hub',           'High',   90.5, 16],
        ['PT-021', 'Boston Medical Center',        'Medium', 77.0, 27],
        ['PT-022', 'Houston Research Institute',   'Low',    64.5, 41],
        ['PT-023', 'Seattle Cancer Center',        'High',   86.8, 19],
        ['PT-024', 'Chicago Health Network',       'Medium', 73.4, 34],
        ['PT-025', 'Miami Clinical Hub',           'Low',    57.6, 39],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 12
    ws1.column_dimensions['B'].width = 30
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 26

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet2: Summary (empty — agent must build pivot tables here) ---
    ws2 = wb.create_sheet('Summary')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
