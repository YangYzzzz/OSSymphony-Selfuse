"""
Initial Setup: Class Schedule sorted alphabetically by Day (not weekly order)
Task ID: calc_dop_sort_custom_073
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_sort_custom_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Schedule ---
    ws = wb.active
    ws.title = 'Schedule'

    # Headers
    headers = ['Day', 'Time', 'Class', 'Instructor', 'Room', 'Capacity']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data: 28 rows, sorted ALPHABETICALLY by Day (Friday, Monday, Saturday, Sunday, Thursday, Tuesday, Wednesday)
    # This is the initial state BEFORE the agent sorts by custom order
    data = [
        # Friday (5 rows)
        ['Friday', '08:00', 'Yoga Flow', 'Amanda Torres', 'Studio A', 20],
        ['Friday', '10:00', 'Pilates Core', 'Steven Park', 'Studio B', 15],
        ['Friday', '12:00', 'Spin Cycle', 'Rebecca Hill', 'Cycling Room', 25],
        ['Friday', '14:00', 'Zumba Dance', 'Carlos Mendez', 'Studio A', 30],
        ['Friday', '17:00', 'Body Pump', 'Diana Walsh', 'Main Hall', 35],
        # Monday (5 rows)
        ['Monday', '07:00', 'Morning Stretch', 'Laura Kim', 'Studio A', 18],
        ['Monday', '09:00', 'Kickboxing', 'Marcus Johnson', 'Main Hall', 25],
        ['Monday', '11:00', 'HIIT Training', 'Sarah Chen', 'Studio B', 20],
        ['Monday', '14:00', 'Aqua Aerobics', 'Tom Rivera', 'Pool', 15],
        ['Monday', '18:00', 'Evening Yoga', 'Priya Patel', 'Studio A', 22],
        # Saturday (4 rows)
        ['Saturday', '09:00', 'Boot Camp', 'Kevin O\'Brien', 'Main Hall', 30],
        ['Saturday', '10:30', 'Dance Cardio', 'Jennifer Lee', 'Studio A', 25],
        ['Saturday', '12:00', 'Strength Circuit', 'Alex Nguyen', 'Weight Room', 15],
        ['Saturday', '14:00', 'Tai Chi', 'Wei Zhang', 'Studio B', 20],
        # Sunday (3 rows)
        ['Sunday', '10:00', 'Gentle Yoga', 'Priya Patel', 'Studio A', 20],
        ['Sunday', '11:30', 'Family Swim', 'Tom Rivera', 'Pool', 30],
        ['Sunday', '13:00', 'Meditation', 'Laura Kim', 'Studio B', 15],
        # Thursday (4 rows)
        ['Thursday', '08:00', 'Power Yoga', 'Amanda Torres', 'Studio A', 18],
        ['Thursday', '10:00', 'Aerobics', 'Diana Walsh', 'Main Hall', 30],
        ['Thursday', '13:00', 'Core Stability', 'Sarah Chen', 'Studio B', 20],
        ['Thursday', '16:00', 'Indoor Cycling', 'Rebecca Hill', 'Cycling Room', 22],
        # Tuesday (4 rows)
        ['Tuesday', '07:30', 'Sunrise Run', 'Marcus Johnson', 'Track', 20],
        ['Tuesday', '09:00', 'Barre Fitness', 'Jennifer Lee', 'Studio A', 16],
        ['Tuesday', '11:00', 'Swim Lessons', 'Tom Rivera', 'Pool', 12],
        ['Tuesday', '15:00', 'Boxing Basics', 'Kevin O\'Brien', 'Main Hall', 18],
        # Wednesday (3 rows)
        ['Wednesday', '08:30', 'Vinyasa Yoga', 'Priya Patel', 'Studio A', 20],
        ['Wednesday', '10:00', 'CrossFit', 'Alex Nguyen', 'Weight Room', 15],
        ['Wednesday', '12:00', 'Water Polo', 'Tom Rivera', 'Pool', 18],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total data rows: {len(data)}')
    print('Data sorted alphabetically by Day (Friday first - NOT weekly order)')


create_initial()
