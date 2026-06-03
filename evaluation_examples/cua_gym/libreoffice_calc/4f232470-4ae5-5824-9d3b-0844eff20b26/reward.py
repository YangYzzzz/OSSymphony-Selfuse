"""
Reward Script: Sort class schedule by Day of Week using custom order
Task ID: calc_dop_sort_custom_073
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Monday classes appear first (rows 2-6 all have Day=Monday)
  Component 2 (0.4): All days follow custom weekly order Mon->Tue->Wed->Thu->Fri->Sat->Sun throughout the file
  Component 3 (0.2): Data integrity — all 28 original records remain present with correct row pairing
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_sort_custom_073'

# Custom day order for the sort task
DAY_ORDER = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

# Expected full dataset in golden order (each entry is [Day, Time, Class, Instructor, Room, Capacity])
EXPECTED_ROWS = [
    ['Monday', '07:00', 'Morning Stretch', 'Laura Kim', 'Studio A', 18],
    ['Monday', '09:00', 'Kickboxing', 'Marcus Johnson', 'Main Hall', 25],
    ['Monday', '11:00', 'HIIT Training', 'Sarah Chen', 'Studio B', 20],
    ['Monday', '14:00', 'Aqua Aerobics', 'Tom Rivera', 'Pool', 15],
    ['Monday', '18:00', 'Evening Yoga', 'Priya Patel', 'Studio A', 22],
    ['Tuesday', '07:30', 'Sunrise Run', 'Marcus Johnson', 'Track', 20],
    ['Tuesday', '09:00', 'Barre Fitness', 'Jennifer Lee', 'Studio A', 16],
    ['Tuesday', '11:00', 'Swim Lessons', 'Tom Rivera', 'Pool', 12],
    ['Tuesday', '15:00', 'Boxing Basics', "Kevin O'Brien", 'Main Hall', 18],
    ['Wednesday', '08:30', 'Vinyasa Yoga', 'Priya Patel', 'Studio A', 20],
    ['Wednesday', '10:00', 'CrossFit', 'Alex Nguyen', 'Weight Room', 15],
    ['Wednesday', '12:00', 'Water Polo', 'Tom Rivera', 'Pool', 18],
    ['Thursday', '08:00', 'Power Yoga', 'Amanda Torres', 'Studio A', 18],
    ['Thursday', '10:00', 'Aerobics', 'Diana Walsh', 'Main Hall', 30],
    ['Thursday', '13:00', 'Core Stability', 'Sarah Chen', 'Studio B', 20],
    ['Thursday', '16:00', 'Indoor Cycling', 'Rebecca Hill', 'Cycling Room', 22],
    ['Friday', '08:00', 'Yoga Flow', 'Amanda Torres', 'Studio A', 20],
    ['Friday', '10:00', 'Pilates Core', 'Steven Park', 'Studio B', 15],
    ['Friday', '12:00', 'Spin Cycle', 'Rebecca Hill', 'Cycling Room', 25],
    ['Friday', '14:00', 'Zumba Dance', 'Carlos Mendez', 'Studio A', 30],
    ['Friday', '17:00', 'Body Pump', 'Diana Walsh', 'Main Hall', 35],
    ['Saturday', '09:00', 'Boot Camp', "Kevin O'Brien", 'Main Hall', 30],
    ['Saturday', '10:30', 'Dance Cardio', 'Jennifer Lee', 'Studio A', 25],
    ['Saturday', '12:00', 'Strength Circuit', 'Alex Nguyen', 'Weight Room', 15],
    ['Saturday', '14:00', 'Tai Chi', 'Wei Zhang', 'Studio B', 20],
    ['Sunday', '10:00', 'Gentle Yoga', 'Priya Patel', 'Studio A', 20],
    ['Sunday', '11:30', 'Family Swim', 'Tom Rivera', 'Pool', 30],
    ['Sunday', '13:00', 'Meditation', 'Laura Kim', 'Studio B', 15],
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task is to sort the class schedule by Day of Week using a custom order:
    Monday, Tuesday, Wednesday, Thursday, Friday, Saturday, Sunday
    (not the default alphabetical order).
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify the Schedule sheet exists
    if 'Schedule' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Schedule' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Schedule']

    # Precondition: verify headers are intact
    expected_headers = ['Day', 'Time', 'Class', 'Instructor', 'Room', 'Capacity']
    actual_headers = [ws.cell(row=1, column=c).value for c in range(1, 7)]
    if actual_headers != expected_headers:
        print(f"CRITICAL: Headers not intact. Expected {expected_headers}, found {actual_headers}")
        print("REWARD: 0.0")
        return 0.0

    # Read all data rows
    data_rows = []
    for r in range(2, 30):
        row = [ws.cell(row=r, column=c).value for c in range(1, 7)]
        data_rows.append(row)

    # -------------------------------------------------------------------------
    # Component 1: Monday classes appear first (0.4 points)
    # In the golden file, rows 2-6 should all have Day = 'Monday'.
    # This FAILS on the initial file (which starts with Friday) and
    # PASSES on the golden file.
    # -------------------------------------------------------------------------
    try:
        # The first 5 data entries should all be Monday
        first_five_days = [row[0] for row in data_rows[:5]]
        all_monday = all(day == 'Monday' for day in first_five_days)
        if all_monday:
            print(f"PASS: Component 1 — Monday classes appear first (rows 2-6 all have Day='Monday'). (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Expected first 5 rows to be Monday, found: {first_five_days}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: All days follow custom weekly order Mon->Tue->Wed->Thu->Fri->Sat->Sun (0.4 points)
    # In the golden file, the day column (A) must be non-decreasing in the
    # custom order throughout all 28 data rows.
    # This FAILS on the initial file (which has alphabetical order) and
    # PASSES on the golden file.
    # -------------------------------------------------------------------------
    try:
        days_in_file = [row[0] for row in data_rows]
        violation = None
        prev_day_idx = -1
        for i, day in enumerate(days_in_file):
            if day not in DAY_ORDER:
                violation = f"Unknown day '{day}' at row {i+2}"
                break
            idx = DAY_ORDER.index(day)
            if idx < prev_day_idx:
                violation = f"Order violation at row {i+2}: '{day}' (index {idx}) follows day at index {prev_day_idx}"
                break
            prev_day_idx = idx

        if violation is None:
            print(f"PASS: Component 2 — All 28 rows follow custom order Mon->Tue->Wed->Thu->Fri->Sat->Sun. (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Custom order not followed. {violation}")
            print(f"  Day sequence found: {days_in_file}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Data integrity — all 28 original records present, row data intact (0.2 points)
    # Verifies that no rows were lost or corrupted during the sort operation.
    # Checks that the exact set of records from the initial file still exist in the golden file.
    # NOTE: This checks that row data is paired correctly (each row's class/instructor/room stay together).
    # This FAILS on the initial file (which has different row order, so exact match against
    # EXPECTED_ROWS fails) and PASSES on the golden file.
    # -------------------------------------------------------------------------
    try:
        if len(data_rows) != 28:
            print(f"FAIL: Component 3 — Expected 28 data rows, found {len(data_rows)}")
        else:
            # Check each row matches the expected golden ordering
            mismatches = []
            for i, (actual, expected) in enumerate(zip(data_rows, EXPECTED_ROWS)):
                if actual != expected:
                    mismatches.append(f"Row {i+2}: expected {expected}, found {actual}")

            if not mismatches:
                print(f"PASS: Component 3 — All 28 rows match expected data with correct pairing. (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — {len(mismatches)} row(s) do not match expected data.")
                for m in mismatches[:3]:
                    print(f"  {m}")
                if len(mismatches) > 3:
                    print(f"  ... and {len(mismatches) - 3} more")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
