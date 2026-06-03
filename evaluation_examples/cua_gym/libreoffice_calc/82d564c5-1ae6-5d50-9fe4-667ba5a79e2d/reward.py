"""
Reward Script: Create pivot table with 'Total Domestic' calculated item
Task ID: calc_gcp_064
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): Pivot sheet exists with correct headers
  Component 2 (0.3): Individual region rows with SUM of Amount values
  Component 3 (0.3): 'Total Domestic' calculated item = sum of 4 domestic regions
  Component 4 (0.2): Grand Total row and overall structure
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_064'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---- Precondition: RegionalSales sheet must still exist ----
    if 'RegionalSales' not in wb.sheetnames:
        print("FAIL: RegionalSales sheet missing — data integrity compromised")
        print("REWARD: 0.0")
        return 0.0

    # ---- Component 1: Pivot sheet exists with correct headers (0.2 pts) ----
    # This FAILS on initial (no Pivot sheet) and PASSES on golden
    try:
        pivot_sheet = None
        # Look for a sheet that serves as the pivot table output
        # Accept any sheet name that isn't 'RegionalSales'
        for sn in wb.sheetnames:
            if sn != 'RegionalSales':
                ws_candidate = wb[sn]
                # Check if it has pivot-like structure: Region header + Amount header
                header_a = ws_candidate.cell(row=1, column=1).value
                header_b = ws_candidate.cell(row=1, column=2).value
                if header_a and header_b:
                    header_a_str = str(header_a).lower().strip()
                    header_b_str = str(header_b).lower().strip()
                    if 'region' in header_a_str and ('amount' in header_b_str or 'sum' in header_b_str):
                        pivot_sheet = ws_candidate
                        break

        if pivot_sheet is not None:
            print(f"PASS: Component 1 — Pivot sheet '{pivot_sheet.title}' found with headers "
                  f"'{pivot_sheet.cell(row=1, column=1).value}' / '{pivot_sheet.cell(row=1, column=2).value}' (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — No pivot-like sheet found with Region + Amount headers")
            print(f"  Available sheets: {wb.sheetnames}")
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # ---- Build a map of region -> amount from the pivot sheet ----
    region_map = {}
    for r in range(2, pivot_sheet.max_row + 1):
        label = pivot_sheet.cell(row=r, column=1).value
        value = pivot_sheet.cell(row=r, column=2).value
        if label is not None:
            region_map[str(label).strip()] = value

    print(f"  Pivot data found: {region_map}")

    # ---- Component 2: Individual region rows present with numeric values (0.3 pts) ----
    # The task requires North, South, East, West, International as individual rows
    # This FAILS on initial (no pivot sheet) and PASSES on golden
    try:
        required_regions = ['North', 'South', 'East', 'West', 'International']
        found_regions = []
        for reg in required_regions:
            # Case-insensitive match
            matched = None
            for key in region_map:
                if key.lower() == reg.lower():
                    matched = key
                    break
            if matched and isinstance(region_map[matched], (int, float)) and region_map[matched] > 0:
                found_regions.append(matched)

        region_ratio = len(found_regions) / len(required_regions)
        region_score = 0.3 * region_ratio

        if region_ratio == 1.0:
            print(f"PASS: Component 2 — All 5 regions found with numeric values ({0.3} pts)")
            total_score += 0.3
        elif region_ratio > 0:
            print(f"PARTIAL: Component 2 — {len(found_regions)}/5 regions found: {found_regions} ({region_score:.2f} pts)")
            total_score += region_score
        else:
            print(f"FAIL: Component 2 — No required regions found. Available keys: {list(region_map.keys())}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---- Component 3: 'Total Domestic' calculated item (0.3 pts) ----
    # This is THE key task requirement: a row named 'Total Domestic' that sums North+South+East+West
    # This FAILS on initial (no pivot sheet, no such row) and PASSES on golden
    try:
        # Find Total Domestic row (case-insensitive, flexible naming)
        td_value = None
        td_key = None
        for key in region_map:
            if 'total' in key.lower() and 'domestic' in key.lower():
                td_value = region_map[key]
                td_key = key
                break

        if td_key is None:
            print("FAIL: Component 3 — No 'Total Domestic' row found in pivot")
        elif not isinstance(td_value, (int, float)):
            print(f"FAIL: Component 3 — 'Total Domestic' value is not numeric: {td_value}")
        else:
            # Calculate expected domestic sum from individual regions
            domestic_regions = ['North', 'South', 'East', 'West']
            domestic_sum = 0.0
            domestic_found_count = 0
            for reg in domestic_regions:
                matched = None
                for key in region_map:
                    if key.lower() == reg.lower():
                        matched = key
                        break
                if matched and isinstance(region_map[matched], (int, float)):
                    domestic_sum += region_map[matched]
                    domestic_found_count += 1

            if domestic_found_count == 4 and abs(td_value - domestic_sum) < 1.0:
                print(f"PASS: Component 3 — 'Total Domestic' = {td_value}, expected sum = {domestic_sum} (0.3 pts)")
                total_score += 0.3
            elif domestic_found_count == 4:
                print(f"FAIL: Component 3 — 'Total Domestic' = {td_value}, but expected ~{domestic_sum} "
                      f"(diff={abs(td_value - domestic_sum):.2f})")
            elif domestic_found_count > 0:
                # Partial: Total Domestic exists but can't fully verify sum
                if td_value > 0:
                    print(f"PARTIAL: Component 3 — 'Total Domestic' = {td_value} but only {domestic_found_count}/4 domestic regions found (0.1 pts)")
                    total_score += 0.1
            else:
                print(f"FAIL: Component 3 — 'Total Domestic' exists but no domestic region values found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---- Component 4: Grand Total row and structure (0.2 pts) ----
    # Pivot should have a Grand Total row. This FAILS on initial (no pivot) and PASSES on golden
    try:
        gt_value = None
        gt_key = None
        for key in region_map:
            if 'grand' in key.lower() and 'total' in key.lower():
                gt_value = region_map[key]
                gt_key = key
                break

        if gt_key is None:
            print("FAIL: Component 4 — No 'Grand Total' row found")
        elif not isinstance(gt_value, (int, float)):
            print(f"FAIL: Component 4 — 'Grand Total' value is not numeric: {gt_value}")
        else:
            # Verify Grand Total = Total Domestic + International (or sum of all)
            intl_val = None
            for key in region_map:
                if key.lower() == 'international':
                    intl_val = region_map[key]
                    break

            # Also try: Grand Total should equal sum of all individual region values
            all_individual_sum = 0.0
            individual_regions = ['North', 'South', 'East', 'West', 'International']
            for reg in individual_regions:
                for key in region_map:
                    if key.lower() == reg.lower():
                        if isinstance(region_map[key], (int, float)):
                            all_individual_sum += region_map[key]
                        break

            if abs(gt_value - all_individual_sum) < 1.0:
                print(f"PASS: Component 4 — Grand Total = {gt_value}, sum of 5 regions = {all_individual_sum} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 — Grand Total = {gt_value}, expected ~{all_individual_sum} (diff={abs(gt_value - all_individual_sum):.2f})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
