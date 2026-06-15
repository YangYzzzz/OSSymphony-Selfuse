"""
Initial Setup: Regional sales data for pivot table task
Task ID: calc_gcp_064
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RegionalSales'

    # Headers
    headers = ['SaleID', 'Date', 'Region', 'Product', 'Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Region distribution: ~20% each for North/South/East/West, ~20% International
    domestic_regions = ['North', 'South', 'East', 'West']
    regions_weighted = domestic_regions * 4 + ['International'] * 4  # 80% domestic, 20% intl

    products = [
        'Widget Pro', 'DataSync License', 'CloudVault Subscription',
        'PrintMaster Ink', 'SecureNet VPN', 'AnalyticsPro Suite',
        'ServerRack Unit', 'DesktopGuard Antivirus', 'SmartBoard Display',
        'NetSwitch Router', 'BackupDrive 2TB', 'CodeEditor Enterprise'
    ]

    # Generate 500 rows of data
    for i in range(1, 501):
        sale_id = i
        # Random date in 2025
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        date_str = f'2025-{month:02d}-{day:02d}'
        region = random.choice(regions_weighted)
        product = random.choice(products)

        # Amount ranges differ slightly by region for realism
        if region == 'International':
            amount = round(random.uniform(150, 2500), 2)
        elif region == 'North':
            amount = round(random.uniform(100, 2200), 2)
        elif region == 'South':
            amount = round(random.uniform(80, 1800), 2)
        elif region == 'East':
            amount = round(random.uniform(120, 2000), 2)
        else:  # West
            amount = round(random.uniform(90, 2100), 2)

        row = i + 1
        ws.cell(row=row, column=1, value=sale_id)
        ws.cell(row=row, column=2, value=date_str)
        ws.cell(row=row, column=3, value=region)
        ws.cell(row=row, column=4, value=product)
        ws.cell(row=row, column=5, value=amount)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 28
    ws.column_dimensions['E'].width = 14

    # Format header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Number format for Amount column
    for row in range(2, 502):
        ws.cell(row=row, column=5).number_format = '#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
