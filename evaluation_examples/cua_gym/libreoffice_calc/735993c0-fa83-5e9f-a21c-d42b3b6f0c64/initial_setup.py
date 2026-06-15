"""
Initial Setup: Define print area task - create spreadsheet with main data and helper columns
Task ID: calc_gfl_051
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # --- Main data columns A-G (headers) ---
    headers_main = ["ID", "Employee", "Department", "Q1 Revenue", "Q2 Revenue", "Q3 Revenue", "Q4 Revenue"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers_main, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Helper columns H-M (headers) ---
    headers_helper = ["YTD Total", "Avg Quarterly", "Growth %", "Rank", "Tier", "Notes Flag"]
    for col, h in enumerate(headers_helper, 8):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=10, italic=True, color="808080")

    # --- Data rows 2-35 (34 employees) ---
    departments = ["Engineering", "Marketing", "Sales", "Finance", "Operations", "HR", "Legal", "Support"]
    tiers = ["Gold", "Silver", "Bronze"]

    employees = [
        ("Sarah Chen", "Engineering", 48200, 51300, 49800, 53100),
        ("Marcus Johnson", "Marketing", 32100, 34500, 36200, 35800),
        ("Priya Patel", "Sales", 61200, 58900, 63400, 67200),
        ("David Kim", "Engineering", 45600, 47200, 46800, 49100),
        ("Elena Rodriguez", "Finance", 38900, 39200, 40100, 41500),
        ("James O'Brien", "Operations", 29800, 31200, 30500, 32100),
        ("Aisha Mohammed", "HR", 27500, 28100, 29300, 28800),
        ("Thomas Weber", "Sales", 55300, 57800, 54200, 59100),
        ("Lisa Nakamura", "Engineering", 51200, 53400, 52100, 54800),
        ("Carlos Mendez", "Marketing", 33800, 35100, 34600, 36900),
        ("Rachel Foster", "Legal", 42100, 43500, 44200, 45100),
        ("Wei Zhang", "Finance", 37200, 38100, 39500, 40200),
        ("Sophie Martin", "Support", 25600, 26800, 27100, 28300),
        ("Ahmed Hassan", "Operations", 31500, 32800, 33100, 34500),
        ("Olivia Brown", "Engineering", 49300, 50100, 51800, 52600),
        ("Ryan Murphy", "Sales", 58100, 56200, 59800, 61400),
        ("Yuki Tanaka", "Marketing", 34200, 35800, 36100, 37500),
        ("Laura Schmidt", "HR", 28200, 29100, 30400, 29800),
        ("Daniel Park", "Finance", 39100, 40500, 41200, 42800),
        ("Maria Gonzalez", "Support", 26100, 27300, 28500, 29100),
        ("Christopher Lee", "Engineering", 47800, 49200, 50100, 51500),
        ("Anna Kowalski", "Legal", 43200, 44100, 45500, 46200),
        ("Michael Davis", "Operations", 30200, 31500, 32800, 33400),
        ("Fatima Al-Rashid", "Sales", 56800, 58100, 60200, 62500),
        ("Brian Thompson", "Marketing", 31900, 33200, 34800, 35100),
        ("Ingrid Larsson", "Finance", 40100, 41200, 42500, 43800),
        ("Kevin Wright", "Support", 24800, 25900, 26700, 27500),
        ("Nadia Petrova", "Engineering", 50400, 51800, 53200, 54100),
        ("Samuel Okafor", "HR", 29500, 30200, 31100, 30800),
        ("Julia Fischer", "Operations", 32100, 33500, 34200, 35800),
        ("Antonio Rossi", "Legal", 44500, 45200, 46800, 47100),
        ("Hana Yoshida", "Sales", 57200, 59100, 58400, 61800),
        ("Patrick O'Connor", "Engineering", 46200, 48100, 47500, 50200),
        ("Zara Khalil", "Marketing", 35100, 36400, 37200, 38500),
    ]

    for r, (name, dept, q1, q2, q3, q4) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=r - 1)  # ID
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=dept)
        ws.cell(row=r, column=4, value=q1)
        ws.cell(row=r, column=5, value=q2)
        ws.cell(row=r, column=6, value=q3)
        ws.cell(row=r, column=7, value=q4)

        # Helper columns H-M with intermediate calculations
        ytd = q1 + q2 + q3 + q4
        avg_q = ytd / 4
        growth = round(((q4 - q1) / q1) * 100, 2) if q1 != 0 else 0
        rank = r - 1
        tier = tiers[0] if ytd > 200000 else (tiers[1] if ytd > 150000 else tiers[2])
        notes_flag = 1 if growth > 10 else 0

        ws.cell(row=r, column=8, value=ytd)
        ws.cell(row=r, column=9, value=round(avg_q, 2))
        ws.cell(row=r, column=10, value=growth)
        ws.cell(row=r, column=11, value=rank)
        ws.cell(row=r, column=12, value=tier)
        ws.cell(row=r, column=13, value=notes_flag)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 10
    ws.column_dimensions["K"].width = 8
    ws.column_dimensions["L"].width = 8
    ws.column_dimensions["M"].width = 12

    # Number formatting for currency columns
    for r in range(2, 36):
        for c in [4, 5, 6, 7, 8, 9]:
            ws.cell(row=r, column=c).number_format = '#,##0'
        ws.cell(row=r, column=10).number_format = '0.00"%"'

    # NO print area set - that's the task for the agent
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
