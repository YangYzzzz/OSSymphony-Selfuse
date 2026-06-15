"""
Reward Script: Set up the 2025 company holiday calendar
Task ID: calc_hr_holiday_calendar_067
Domain: libreoffice_calc
Scoring:
  - Component 1: Column B (B2:B13) contains TEXT formula for day of week (0.3 pts)
  - Component 2: Column D (D2:D13) contains IF/NETWORKDAYS formula for working days (0.3 pts)
  - Component 3: Conditional formatting rule for past holidays (grey #D9D9D9) (0.2 pts)
  - Component 4: Conditional formatting rule for upcoming within 30 days (green #70AD47) (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_holiday_calendar_067'
SHEET_NAME = '2025 Holidays'


def normalize_formula(formula):
    """Normalize formula for comparison: strip whitespace and uppercase."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def check_text_formula_in_col_b(ws):
    """
    Check that all 12 rows in column B contain =TEXT(Ax,"DDDD") formulas.
    Returns count of valid formulas found (max 12).
    """
    count = 0
    fail_details = []
    for row in range(2, 14):
        val = ws.cell(row=row, column=2).value
        if val is None:
            fail_details.append(f"B{row}=None")
            continue
        norm = normalize_formula(str(val))
        # Must reference correct row, use TEXT function, and format DDDD
        if 'TEXT' in norm and f'A{row}' in norm and 'DDDD' in norm:
            count += 1
        else:
            fail_details.append(f"B{row}={repr(val)}")
    return count, fail_details


def check_if_networkdays_formula_in_col_d(ws):
    """
    Check that all 12 rows in column D contain IF/TODAY/NETWORKDAYS formulas.
    Returns count of valid formulas found (max 12).
    """
    count = 0
    fail_details = []
    for row in range(2, 14):
        val = ws.cell(row=row, column=4).value
        if val is None:
            fail_details.append(f"D{row}=None")
            continue
        norm = normalize_formula(str(val))
        # Must contain IF, TODAY, NETWORKDAYS, referencing correct Ax row
        if ('IF' in norm and 'TODAY()' in norm and
                'NETWORKDAYS' in norm and f'A{row}' in norm):
            count += 1
        else:
            fail_details.append(f"D{row}={repr(val)}")
    return count, fail_details


def find_cf_rule_grey(ws):
    """
    Search conditional formatting for past holiday rule (grey #D9D9D9).
    Rule pattern: expression with A<TODAY(), fill FFD9D9D9.
    Returns (rule_found, color_correct, range_correct).
    """
    rule_found = False
    color_correct = False
    range_correct = False
    for cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
        range_str = str(cf_range)
        covers_range = ('A2' in range_str and 'D13' in range_str)
        for rule in cf_rules:
            if rule.type != 'expression':
                continue
            if not (hasattr(rule, 'formula') and rule.formula):
                continue
            formula_text = normalize_formula(str(rule.formula[0]))
            # Past holiday rule: references A column and TODAY(), but NOT D column condition
            is_past_rule = ('TODAY()' in formula_text and
                            ('$A' in formula_text or 'A2' in formula_text or 'A3' in formula_text) and
                            'ISNUMBER' not in formula_text and
                            '<=30' not in formula_text)
            if not is_past_rule:
                continue
            rule_found = True
            if covers_range:
                range_correct = True
            # Check grey color
            try:
                fill = rule.dxf.fill
                if fill and hasattr(fill, 'fgColor'):
                    rgb = fill.fgColor.rgb.upper()
                    if rgb in ('FFD9D9D9', 'D9D9D9'):
                        color_correct = True
            except Exception:
                pass
    return rule_found, color_correct, range_correct


def find_cf_rule_green(ws):
    """
    Search conditional formatting for upcoming holiday rule (green #70AD47).
    Rule pattern: expression with D<=30 or ISNUMBER/D<=30, fill FF70AD47.
    Returns (rule_found, color_correct, range_correct).
    """
    rule_found = False
    color_correct = False
    range_correct = False
    for cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
        range_str = str(cf_range)
        covers_range = ('A2' in range_str and 'D13' in range_str)
        for rule in cf_rules:
            if rule.type != 'expression':
                continue
            if not (hasattr(rule, 'formula') and rule.formula):
                continue
            formula_text = normalize_formula(str(rule.formula[0]))
            # Upcoming within 30 days: references D column, <=30
            is_upcoming_rule = ('$D' in formula_text or 'D2' in formula_text) and '<=30' in formula_text
            if not is_upcoming_rule:
                continue
            rule_found = True
            if covers_range:
                range_correct = True
            # Check green color
            try:
                fill = rule.dxf.fill
                if fill and hasattr(fill, 'fgColor'):
                    rgb = fill.fgColor.rgb.upper()
                    if rgb in ('FF70AD47', '70AD47'):
                        color_correct = True
            except Exception:
                pass
    return rule_found, color_correct, range_correct


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — fail early if file cannot be opened
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -------------------------------------------------------------------
    # Component 1: Column B (B2:B13) must contain =TEXT(Ax,"DDDD") formulas
    # These are absent in the initial file (all None), present in golden.
    # -------------------------------------------------------------------
    try:
        b_count, b_fails = check_text_formula_in_col_b(ws)
        b_total = 12
        if b_count == b_total:
            print(f"PASS: Component 1 — All 12 B-column TEXT(Ax,DDDD) formulas present (0.3 pts)")
            total_score += 0.3
        elif b_count >= 6:
            partial = round(0.3 * b_count / b_total, 4)
            print(f"PARTIAL: Component 1 — {b_count}/{b_total} B-column TEXT formulas present ({partial} pts)")
            print(f"  Issues: {b_fails[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {b_count}/{b_total} B-column TEXT formulas found")
            if b_fails:
                print(f"  Issues: {b_fails[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Column D (D2:D13) must contain =IF(Ax<TODAY(),"Past",NETWORKDAYS(TODAY(),Ax))
    # These are absent in the initial file (all None), present in golden.
    # -------------------------------------------------------------------
    try:
        d_count, d_fails = check_if_networkdays_formula_in_col_d(ws)
        d_total = 12
        if d_count == d_total:
            print(f"PASS: Component 2 — All 12 D-column IF/NETWORKDAYS formulas present (0.3 pts)")
            total_score += 0.3
        elif d_count >= 6:
            partial = round(0.3 * d_count / d_total, 4)
            print(f"PARTIAL: Component 2 — {d_count}/{d_total} D-column formulas present ({partial} pts)")
            print(f"  Issues: {d_fails[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {d_count}/{d_total} D-column formulas found")
            if d_fails:
                print(f"  Issues: {d_fails[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Conditional formatting rule for past holidays — grey background #D9D9D9
    # Rule formula: $A2<TODAY(), fill color FFD9D9D9
    # Not present in initial file.
    # -------------------------------------------------------------------
    try:
        grey_found, grey_color_ok, grey_range_ok = find_cf_rule_grey(ws)
        if grey_found and grey_color_ok and grey_range_ok:
            print(f"PASS: Component 3 — Grey CF rule on A2:D13, color FFD9D9D9, formula A<TODAY() (0.2 pts)")
            total_score += 0.2
        elif grey_found and grey_color_ok:
            print(f"PARTIAL: Component 3 — Grey CF rule with correct color but incorrect range (0.1 pts)")
            total_score += 0.1
        elif grey_found:
            print(f"PARTIAL: Component 3 — Grey CF rule found but color or range wrong (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 3 — No grey CF rule for past holidays ($A<TODAY()) found (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: Conditional formatting rule for upcoming within 30 days — green #70AD47
    # Rule formula: AND(ISNUMBER($D2),$D2<=30), fill color FF70AD47
    # Not present in initial file.
    # -------------------------------------------------------------------
    try:
        green_found, green_color_ok, green_range_ok = find_cf_rule_green(ws)
        if green_found and green_color_ok and green_range_ok:
            print(f"PASS: Component 4 — Green CF rule on A2:D13, color FF70AD47, formula D<=30 (0.2 pts)")
            total_score += 0.2
        elif green_found and green_color_ok:
            print(f"PARTIAL: Component 4 — Green CF rule with correct color but incorrect range (0.1 pts)")
            total_score += 0.1
        elif green_found:
            print(f"PARTIAL: Component 4 — Green CF rule found but color or range wrong (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 4 — No green CF rule for upcoming holidays (D<=30) found (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
