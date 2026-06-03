"""
Initial Setup: 2025 Company Holiday Calendar
Task ID: calc_hr_holiday_calendar_067
Domain: libreoffice_calc

Creates a spreadsheet with the 2025 company holiday calendar.
Sheet '2025 Holidays' contains:
- Row 1 headers: Date, Day of Week, Holiday Name, Working Days Away
- Rows 2-13: 12 company holidays with dates in column A and names in column C
- Columns B and D are EMPTY (to be filled by the agent)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_holiday_calendar_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '2025 Holidays'

    # --- Headers ---
    headers = ['Date', 'Day of Week', 'Holiday Name', 'Working Days Away']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center')

    # --- 2025 US-style company holidays (12 holidays) ---
    # Column A: date values, Column B: EMPTY, Column C: holiday name, Column D: EMPTY
    holidays = [
        (date(2025, 1, 1),   "New Year's Day"),
        (date(2025, 1, 20),  "Martin Luther King Jr. Day"),
        (date(2025, 2, 17),  "Presidents' Day"),
        (date(2025, 5, 26),  "Memorial Day"),
        (date(2025, 6, 19),  "Juneteenth"),
        (date(2025, 7, 4),   "Independence Day"),
        (date(2025, 9, 1),   "Labor Day"),
        (date(2025, 10, 13), "Columbus Day"),
        (date(2025, 11, 11), "Veterans Day"),
        (date(2025, 11, 27), "Thanksgiving Day"),
        (date(2025, 11, 28), "Day After Thanksgiving"),
        (date(2025, 12, 25), "Christmas Day"),
    ]

    for r, (holiday_date, holiday_name) in enumerate(holidays, 2):
        # Column A: date value
        cell_a = ws.cell(row=r, column=1, value=holiday_date)
        cell_a.number_format = 'YYYY-MM-DD'
        cell_a.alignment = Alignment(horizontal='center')

        # Column B: EMPTY (Day of Week - agent will fill with TEXT formula)
        ws.cell(row=r, column=2, value=None)

        # Column C: Holiday name
        cell_c = ws.cell(row=r, column=3, value=holiday_name)
        cell_c.alignment = Alignment(horizontal='left')

        # Column D: EMPTY (Working Days Away - agent will fill with NETWORKDAYS formula)
        ws.cell(row=r, column=4, value=None)

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: "2025 Holidays"')
    print(f'  Rows: 1 header + 12 holiday rows')
    print(f'  Columns A (Date) and C (Holiday Name) filled')
    print(f'  Columns B (Day of Week) and D (Working Days Away) are EMPTY')


create_initial()
