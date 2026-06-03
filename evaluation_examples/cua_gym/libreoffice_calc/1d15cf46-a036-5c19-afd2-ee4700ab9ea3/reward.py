"""
Reward Script: Daily Cash Register Reconciliation Sheet
Task ID: calc_grs_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Variance formulas in D7:D12
  Component 2 (0.20): Variance % formulas in E7:E12 + TOTAL row formulas
  Component 3 (0.20): Cash drawer formulas (B19, B21) + Daily Summary formulas (B24, B26, B28)
  Component 4 (0.15): Conditional formatting on variance column
  Component 5 (0.10): Monthly Summary totals/averages formulas (rows 35-36)
  Component 6 (0.15): Chart on Monthly Summary sheet
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_077'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print(f"PERSIST: ctrl+s sent for {domain}")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def is_formula(val):
    """Check if a cell value is a formula string."""
    return isinstance(val, str) and val.startswith('=')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both sheets must exist
    if 'Daily Reconciliation' not in wb.sheetnames:
        print("FAIL: 'Daily Reconciliation' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Daily Reconciliation']

    # Component 1: Variance formulas in D7:D12 (0.20 points)
    # Initial has D7:D12 as None; golden has =B7-C7, etc.
    try:
        variance_formula_count = 0
        for row in range(7, 13):
            val = ws.cell(row=row, column=4).value  # column D
            if is_formula(val):
                variance_formula_count += 1
        if variance_formula_count >= 5:
            print(f"PASS: Component 1 - {variance_formula_count}/6 variance formulas in D7:D12 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 - Only {variance_formula_count}/6 variance formulas in D7:D12 (need >= 5)")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Variance % formulas in E7:E12 + TOTAL row formulas in B13:E13 (0.20 points)
    try:
        pct_formula_count = 0
        for row in range(7, 13):
            val = ws.cell(row=row, column=5).value  # column E
            if is_formula(val):
                pct_formula_count += 1

        total_formula_count = 0
        for col in range(2, 6):  # B13 through E13
            val = ws.cell(row=13, column=col).value
            if is_formula(val):
                total_formula_count += 1

        combined = pct_formula_count + total_formula_count
        # Need at least 5 pct formulas + at least 2 total formulas = 7 minimum
        if pct_formula_count >= 5 and total_formula_count >= 2:
            print(f"PASS: Component 2 - {pct_formula_count} Variance% formulas + {total_formula_count} TOTAL formulas (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 - {pct_formula_count}/6 Variance% formulas, {total_formula_count}/4 TOTAL formulas")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Cash drawer formulas (B19, B21) + Daily Summary formulas (B24, B26, B28) (0.20 points)
    # Initial has these cells as None; golden has formulas
    try:
        formula_cells = {
            'B19': ws.cell(row=19, column=2).value,  # Expected Closing Float
            'B21': ws.cell(row=21, column=2).value,  # Over/Short
            'B24': ws.cell(row=24, column=2).value,  # Net Sales
            'B26': ws.cell(row=26, column=2).value,  # Net Revenue
            'B28': ws.cell(row=28, column=2).value,  # Daily Gross Margin
        }
        formulas_found = sum(1 for v in formula_cells.values() if is_formula(v))
        if formulas_found >= 4:
            print(f"PASS: Component 3 - {formulas_found}/5 cash drawer + daily summary formulas (0.20 pts)")
            total_score += 0.20
        else:
            details = ", ".join(f"{k}={repr(v)}" for k, v in formula_cells.items())
            print(f"FAIL: Component 3 - Only {formulas_found}/5 formulas found: {details}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Conditional formatting on D7:D12 (0.15 points)
    # Initial has 0 conditional formatting rules; golden has rules on D7:D12
    try:
        cf_rules = list(ws.conditional_formatting)
        cf_count = 0
        variance_cf_found = sum(1 for cf in cf_rules if 'D' in str(cf))
        for cf in cf_rules:
            cf_count += len(cf.rules)

        if variance_cf_found > 0 and cf_count >= 2:
            print(f"PASS: Component 4 - Conditional formatting found: {cf_count} rules targeting variance column (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 - Expected CF on variance column, found {cf_count} rules, D-column CF ranges: {variance_cf_found}")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Monthly Summary totals/averages formulas (0.10 points)
    # Initial has no formulas in rows 35-36; golden has SUM/AVERAGE formulas
    try:
        if 'Monthly Summary' not in wb.sheetnames:
            print("FAIL: Component 5 - 'Monthly Summary' sheet not found")
        else:
            ws2 = wb['Monthly Summary']
            summary_formulas = 0
            # Check rows 35 and 36 for SUM/AVERAGE formulas across columns B-E
            for row in range(35, 37):
                for col in range(2, 6):
                    val = ws2.cell(row=row, column=col).value
                    if is_formula(val):
                        summary_formulas += 1

            # Also check if there could be a different row for totals - search nearby
            if summary_formulas < 3:
                # Search rows 34-40 for SUM/AVERAGE formulas
                for row in range(34, 41):
                    for col in range(2, 6):
                        val = ws2.cell(row=row, column=col).value
                        if is_formula(val) and ('SUM' in str(val).upper() or 'AVERAGE' in str(val).upper()):
                            summary_formulas += 1

            if summary_formulas >= 3:
                print(f"PASS: Component 5 - {summary_formulas} summary formulas in Monthly Summary (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 - Only {summary_formulas} summary formulas found in Monthly Summary")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Chart on Monthly Summary sheet (0.15 points)
    # Initial has 0 charts; golden has a LineChart "30-Day Variance Trend"
    try:
        if 'Monthly Summary' not in wb.sheetnames:
            print("FAIL: Component 6 - 'Monthly Summary' sheet not found")
        else:
            ws2 = wb['Monthly Summary']
            chart_count = len(ws2._charts)
            if chart_count >= 1:
                print(f"PASS: Component 6 - {chart_count} chart(s) found on Monthly Summary (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 6 - No charts found on Monthly Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
