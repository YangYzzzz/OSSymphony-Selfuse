"""
Initial Setup: Retail store daily cash register reconciliation sheet
Task ID: calc_grs_077
Domain: libreoffice_calc

Creates a workbook with raw reconciliation data (no formulas, no charts,
no conditional formatting). The agent must add those.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: Daily Reconciliation ==========
    ws1 = wb.active
    ws1.title = "Daily Reconciliation"

    # --- Title area ---
    ws1["A1"] = "Treasure Trove Gift Shop"
    ws1["A1"].font = Font(name="Arial", size=14, bold=True)
    ws1["A2"] = "Daily Cash Register Reconciliation"
    ws1["A2"].font = Font(name="Arial", size=11, bold=True)
    ws1["A3"] = "Date:"
    ws1["B3"] = date(2026, 3, 31).isoformat()

    # --- Payment Types Section ---
    ws1["A5"] = "PAYMENT RECONCILIATION"
    ws1["A5"].font = Font(name="Arial", size=11, bold=True)

    headers_pay = ["Payment Type", "Expected Sales", "Actual Counted",
                   "Variance", "Variance %", "Notes"]
    for c, h in enumerate(headers_pay, 1):
        cell = ws1.cell(row=6, column=c, value=h)
        cell.font = Font(name="Arial", bold=True)

    # Payment type data rows (raw data, NO formulas in Variance/Variance%)
    payment_data = [
        ["Cash",       1245.50, 1238.75, None, None, ""],
        ["Visa",       2387.00, 2387.00, None, None, ""],
        ["Mastercard", 1563.25, 1560.00, None, None, ""],
        ["Amex",        892.00,  892.00, None, None, ""],
        ["EFTPOS",     1734.80, 1734.80, None, None, ""],
        ["Gift Cards",  456.00,  449.50, None, None, ""],
    ]
    for r, row_data in enumerate(payment_data, 7):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                ws1.cell(row=r, column=c, value=val)

    # Totals row placeholder (no formulas)
    ws1.cell(row=13, column=1, value="TOTAL")
    ws1.cell(row=13, column=1).font = Font(name="Arial", bold=True)

    # --- Cash Drawer Section ---
    ws1["A15"] = "CASH DRAWER"
    ws1["A15"].font = Font(name="Arial", size=11, bold=True)

    cash_drawer_labels = [
        ("Opening Float", 200.00),
        ("Total Cash Sales", 1245.50),
        ("Total Cash Refunds", 32.00),
        ("Expected Closing Float", None),  # formula needed
        ("Actual Closing Float", 1407.25),
        ("Over/Short", None),              # formula needed
    ]
    for i, (label, val) in enumerate(cash_drawer_labels):
        row = 16 + i
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=1).font = Font(name="Arial")
        if val is not None:
            ws1.cell(row=row, column=2, value=val)

    # --- Daily Summary Section ---
    ws1["A23"] = "DAILY SUMMARY"
    ws1["A23"].font = Font(name="Arial", size=11, bold=True)

    summary_labels = [
        ("Net Sales", None),          # formula needed
        ("Total Refunds", 32.00),
        ("Net Revenue", None),        # formula needed
        ("Cost of Goods Sold", 3254.80),
        ("Daily Gross Margin", None), # formula needed
    ]
    for i, (label, val) in enumerate(summary_labels):
        row = 24 + i
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=1).font = Font(name="Arial")
        if val is not None:
            ws1.cell(row=row, column=2, value=val)

    # Column widths
    ws1.column_dimensions["A"].width = 24
    ws1.column_dimensions["B"].width = 16
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 14
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 24

    # Number formats for currency columns
    for r in range(7, 13):
        for c in [2, 3, 4]:
            cell = ws1.cell(row=r, column=c)
            if cell.value is not None:
                cell.number_format = '$#,##0.00'

    for r in range(16, 22):
        cell = ws1.cell(row=r, column=2)
        if cell.value is not None:
            cell.number_format = '$#,##0.00'

    for r in range(24, 29):
        cell = ws1.cell(row=r, column=2)
        if cell.value is not None:
            cell.number_format = '$#,##0.00'

    # ========== Sheet 2: Monthly Summary ==========
    ws2 = wb.create_sheet("Monthly Summary")

    ws2["A1"] = "Monthly Reconciliation Summary - March 2026"
    ws2["A1"].font = Font(name="Arial", size=12, bold=True)

    ms_headers = ["Date", "Total Expected", "Total Actual", "Total Variance",
                  "Cash Over/Short", "Notes"]
    for c, h in enumerate(ms_headers, 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", bold=True)

    # 30 days of daily data (realistic gift shop numbers)
    import random
    random.seed(77)  # deterministic

    daily_data = []
    start_date = date(2026, 3, 1)
    for day in range(30):
        d = start_date + timedelta(days=day)
        weekday = d.weekday()
        # Lower sales on Mon/Tue, higher on weekends
        base = 6500 if weekday < 2 else (8500 if weekday >= 5 else 7500)
        expected = round(base + random.uniform(-800, 800), 2)
        # Small variance, occasionally larger
        if random.random() < 0.15:
            variance = round(random.uniform(-25, -8), 2)
        elif random.random() < 0.1:
            variance = round(random.uniform(8, 20), 2)
        else:
            variance = round(random.uniform(-6, 6), 2)
        actual = round(expected - variance, 2)
        cash_short = round(random.uniform(-8, 4), 2)
        note = ""
        if abs(variance) > 10:
            note = "Reviewed"
        daily_data.append([d.isoformat(), expected, actual, variance,
                           cash_short, note])

    for r, row_data in enumerate(daily_data, 4):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # Currency formatting
    for r in range(4, 34):
        for c in [2, 3, 4, 5]:
            ws2.cell(row=r, column=c).number_format = '$#,##0.00'

    # Monthly totals row placeholder (no formulas)
    ws2.cell(row=35, column=1, value="MONTHLY TOTALS")
    ws2.cell(row=35, column=1).font = Font(name="Arial", bold=True)

    # Column widths
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16
    ws2.column_dimensions["E"].width = 16
    ws2.column_dimensions["F"].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
