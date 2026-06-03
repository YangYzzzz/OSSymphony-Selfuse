"""
Initial Setup: Top-grossing animated films spreadsheet with empty worldwide gross column,
plus an empty docx for the agent to write the lowest-grossing film result.
Task ID: osworld_multi_apps_book_reading_rate_015
Domain: multi_apps (libreoffice_calc + libreoffice_writer)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from docx import Document

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_book_reading_rate_015'
XLSX_OUTPUT = f'{WORKDIR}/animated_films.xlsx'
DOCX_OUTPUT = f'{WORKDIR}/lowest_grossing_animated.docx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # --- Create animated_films.xlsx ---
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Animated Films"

    # Headers
    headers = ['Film Title', 'Studio', 'Year', 'Worldwide Gross (USD)']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows — Film Title, Studio, Year filled; Worldwide Gross is intentionally EMPTY
    # These are the top-grossing animated films from 2020-2023 the agent must research
    film_data = [
        ['The Bad Guys',                    'DreamWorks Animation / Universal Pictures', 2022, None],
        ['Turning Red',                      'Pixar / Walt Disney Pictures',             2022, None],
        ['Puss in Boots: The Last Wish',     'DreamWorks Animation / Universal Pictures', 2022, None],
        ['Encanto',                          'Walt Disney Animation Studios',            2021, None],
        ['The Mitchells vs. the Machines',   'Sony Pictures Animation / Netflix',        2021, None],
    ]

    for r, row_data in enumerate(film_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Style the header row (bold)
    from openpyxl.styles import Font
    for col in range(1, 5):
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 42
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 22

    wb.save(XLSX_OUTPUT)
    print(f'Initial spreadsheet created: {XLSX_OUTPUT}')

    # --- Create empty lowest_grossing_animated.docx ---
    doc = Document()
    # Remove default empty paragraph so document is truly empty
    # (keep one empty paragraph so LibreOffice can open it cleanly)
    doc.save(DOCX_OUTPUT)
    print(f'Empty docx created: {DOCX_OUTPUT}')

    # GUI-ready startup: open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{XLSX_OUTPUT}"', delay_sec=2.0)
    # Also open the empty docx in LibreOffice Writer
    launch_gui(f'libreoffice --writer "{DOCX_OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and Writer with DISPLAY=:0')


create_initial()
