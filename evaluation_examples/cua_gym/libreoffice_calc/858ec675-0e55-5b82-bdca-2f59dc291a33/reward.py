"""
Reward Script: Reorder laboratory results table columns to standard reporting format
Task ID: osworld_calc_reorder_columns_007
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column headers in correct order — 0.6 points
    All 8 column headers must be in the target order:
    [Sample ID, Collection Date, Test Type, Analyst, Result Value, Unit, Reference Range, Status]
  Component 2: Data integrity — data rows map correctly to reordered headers — 0.4 points
    Spot-check that Sample ID values are in column 1 and Status values are in column 8,
    confirming the data moved with the headers.

Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_reorder_columns_007'

# Target column order after reordering
TARGET_HEADERS = [
    'Sample ID',
    'Collection Date',
    'Test Type',
    'Analyst',
    'Result Value',
    'Unit',
    'Reference Range',
    'Status'
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook — fail fast if unreadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get active sheet (should be 'Lab Results')
    try:
        if 'Lab Results' in wb.sheetnames:
            ws = wb['Lab Results']
        else:
            ws = wb.active
        print(f"INFO: Using sheet '{ws.title}'")
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must have at least 1 row and 8 columns
    if ws.max_row < 1 or ws.max_column < 8:
        print(f"CRITICAL: Sheet has insufficient size (rows={ws.max_row}, cols={ws.max_column})")
        print("REWARD: 0.0")
        return 0.0

    # --- Component 1: Column headers in correct target order (0.6 points) ---
    # This FAILS on initial_env (wrong order) and PASSES on golden_env (correct order)
    try:
        actual_headers = []
        for col in range(1, 9):
            cell_val = ws.cell(row=1, column=col).value
            actual_headers.append(str(cell_val).strip() if cell_val is not None else '')

        print(f"INFO: Actual headers: {actual_headers}")
        print(f"INFO: Target headers: {TARGET_HEADERS}")

        if actual_headers == TARGET_HEADERS:
            print(f"PASS: Component 1 — All 8 column headers are in the correct target order (0.6 pts)")
            total_score += 0.6
        else:
            # Check how many are in correct position for diagnostic purposes
            correct_count = sum(1 for a, t in zip(actual_headers, TARGET_HEADERS) if a == t)
            print(f"FAIL: Component 1 — Column headers not in correct order. "
                  f"{correct_count}/8 headers in correct position. "
                  f"Expected {TARGET_HEADERS}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check column headers: {e}")

    # --- Component 2: Data integrity — data moved with headers (0.4 points) ---
    # Verify that Sample ID values appear in column 1 and Status values in column 8
    # This FAILS on initial_env (Sample ID is in col 8, Status is in col 2)
    # and PASSES on golden_env (Sample ID in col 1, Status in col 8)
    try:
        error_messages = []
        valid_statuses = {'Normal', 'High', 'Low', 'Critical', 'Abnormal'}

        # Check rows 2 through min(ws.max_row, 5) to spot-check data alignment
        rows_to_check = min(ws.max_row, 5)
        for row_idx in range(2, rows_to_check + 1):
            col1_val = ws.cell(row=row_idx, column=1).value  # Should be Sample ID (e.g., LAB-2025-XXXX)
            col8_val = ws.cell(row=row_idx, column=8).value  # Should be Status (Normal/High/Low)

            # Sample ID check: should start with 'LAB-'
            if col1_val is None or not str(col1_val).startswith('LAB-'):
                error_messages.append(
                    f"Row {row_idx}, Col 1: expected Sample ID starting with 'LAB-', found '{col1_val}'"
                )

            # Status check: should be one of the known status values
            if col8_val is None or str(col8_val).strip() not in valid_statuses:
                error_messages.append(
                    f"Row {row_idx}, Col 8: expected Status (one of {valid_statuses}), found '{col8_val}'"
                )

        if not error_messages:
            print(f"PASS: Component 2 — Data integrity confirmed: "
                  f"Sample ID in col 1 and Status in col 8 for rows 2-{rows_to_check} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Data integrity check failed:")
            for msg in error_messages:
                print(f"  - {msg}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check data integrity: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
