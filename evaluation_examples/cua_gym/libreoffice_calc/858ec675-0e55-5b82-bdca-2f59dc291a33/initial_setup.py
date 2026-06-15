"""
Initial Setup: Laboratory results table with columns in non-standard order
Task ID: osworld_calc_reorder_columns_007
Domain: libreoffice_calc

Creates a spreadsheet with lab results data where columns are in scrambled order:
Result Value, Status, Analyst, Collection Date, Reference Range, Unit, Test Type, Sample ID

The agent needs to reorder to standard format:
Sample ID, Collection Date, Test Type, Analyst, Result Value, Unit, Reference Range, Status
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_reorder_columns_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Lab Results ---
    ws = wb.active
    ws.title = "Lab Results"

    # Columns in SCRAMBLED (non-standard) order:
    # Result Value, Status, Analyst, Collection Date, Reference Range, Unit, Test Type, Sample ID
    headers = [
        "Result Value",
        "Status",
        "Analyst",
        "Collection Date",
        "Reference Range",
        "Unit",
        "Test Type",
        "Sample ID",
    ]

    # Style header row
    header_font = Font(name="Calibri", bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic lab result data
    # Columns: Result Value, Status, Analyst, Collection Date, Reference Range, Unit, Test Type, Sample ID
    data = [
        [7.4,        "Normal",    "Dr. Rachel Kim",      "2025-03-01", "7.35–7.45",   "pH",       "Blood Gas",        "LAB-2025-0001"],
        [142,        "Normal",    "Dr. James Carter",    "2025-03-01", "136–145",      "mmol/L",   "Electrolytes",     "LAB-2025-0002"],
        [3.2,        "Low",       "Dr. Rachel Kim",      "2025-03-02", "3.5–5.1",      "mmol/L",   "Electrolytes",     "LAB-2025-0003"],
        [98,         "Normal",    "Dr. Marcus Lee",      "2025-03-02", "96–106",       "mmol/L",   "Electrolytes",     "LAB-2025-0004"],
        [24.5,       "Normal",    "Dr. James Carter",    "2025-03-03", "22–29",        "mmol/L",   "Electrolytes",     "LAB-2025-0005"],
        [5.8,        "High",      "Dr. Sarah Nguyen",    "2025-03-03", "3.5–5.0",      "mmol/L",   "Glucose",          "LAB-2025-0006"],
        [1.02,       "Normal",    "Dr. Marcus Lee",      "2025-03-04", "1.00–1.30",    "mg/dL",    "Creatinine",       "LAB-2025-0007"],
        [18,         "Normal",    "Dr. Sarah Nguyen",    "2025-03-04", "7–25",         "mg/dL",    "BUN",              "LAB-2025-0008"],
        [4.1,        "Low",       "Dr. Rachel Kim",      "2025-03-05", "4.5–5.9",      "g/dL",     "Hemoglobin",       "LAB-2025-0009"],
        [11200,      "High",      "Dr. James Carter",    "2025-03-05", "4500–11000",   "cells/µL", "WBC Count",        "LAB-2025-0010"],
        [210000,     "Normal",    "Dr. Marcus Lee",      "2025-03-06", "150000–400000","cells/µL", "Platelet Count",   "LAB-2025-0011"],
        [0.8,        "Normal",    "Dr. Sarah Nguyen",    "2025-03-06", "0.0–1.2",      "mg/dL",    "Bilirubin Total",  "LAB-2025-0012"],
        [35,         "Normal",    "Dr. Rachel Kim",      "2025-03-07", "7–40",         "U/L",      "ALT",              "LAB-2025-0013"],
        [42,         "High",      "Dr. James Carter",    "2025-03-07", "10–40",        "U/L",      "AST",              "LAB-2025-0014"],
        [130,        "Low",       "Dr. Marcus Lee",      "2025-03-08", "136–145",      "mmol/L",   "Electrolytes",     "LAB-2025-0015"],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    col_widths = [14, 12, 20, 16, 18, 12, 20, 18]
    for col_idx, width in enumerate(col_widths, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
