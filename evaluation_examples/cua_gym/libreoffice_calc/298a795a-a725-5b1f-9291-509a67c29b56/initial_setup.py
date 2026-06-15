"""
Initial Setup: Build a project dashboard with task dependencies, progress bars,
conditional formatting, and a burndown chart.
Task ID: calc_gpm_014
Domain: libreoffice_calc

Initial state: Plain data only — no formatting, no merged cells, no data validation,
no conditional formatting, no charts, no summary formulas.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProjectDash'

    # Row 1: Plain title (no merge, no formatting)
    ws['A1'] = 'Project Alpha Dashboard'

    # Row 3: Headers (plain, no formatting)
    headers = ['ID', 'Task', 'Depends On', 'Est Hours', 'Actual Hours',
               '% Complete', 'Status', 'Priority', 'Notes']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Rows 4-13: 10 software project tasks with realistic data
    tasks = [
        ['T-001', 'Requirements & Planning', '-', 16, 18, 100,
         'Done', 'High', 'Stakeholder interviews completed'],
        ['T-002', 'UI/UX Design', 'T-001', 24, 22, 100,
         'Done', 'High', 'Figma prototypes approved'],
        ['T-003', 'Database Schema Setup', 'T-001', 12, 10, 100,
         'Done', 'High', 'PostgreSQL with migrations'],
        ['T-004', 'API Development', 'T-003', 40, 35, 85,
         'In Progress', 'High', 'REST endpoints 17/20 done'],
        ['T-005', 'Frontend UI Development', 'T-002', 36, 28, 70,
         'In Progress', 'High', 'React components in progress'],
        ['T-006', 'Integration Layer', 'T-004,T-005', 20, 8, 30,
         'In Progress', 'Medium', 'API-frontend connectors'],
        ['T-007', 'Unit & Integration Testing', 'T-004', 24, 6, 20,
         'In Progress', 'Medium', 'Test coverage target 80%'],
        ['T-008', 'Bug Fixes & QA', 'T-007', 16, 0, 0,
         'To Do', 'Medium', 'Awaiting test results'],
        ['T-009', 'Documentation', 'T-004,T-005', 12, 3, 15,
         'In Progress', 'Low', 'API docs and user guide'],
        ['T-010', 'Deployment & Release', 'T-006,T-007,T-008', 8, 0, 0,
         'Blocked', 'High', 'Blocked on integration testing'],
    ]

    for r, row_data in enumerate(tasks, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # % Complete stored as integer (0-100)

    # Row 15: Summary labels (plain text only, no merge, no bold, no formulas)
    ws['A15'] = 'Summary'
    ws['C15'] = 'Total Hours'

    # Set reasonable column widths so data is visible
    col_widths = {'A': 8, 'B': 28, 'C': 16, 'D': 12, 'E': 14,
                  'F': 12, 'G': 14, 'H': 10, 'I': 32}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
