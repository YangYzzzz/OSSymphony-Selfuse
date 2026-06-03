"""
Reward Script: Build a project dashboard with task dependencies, progress bars,
               conditional formatting, and a burndown chart.
Task ID: calc_gpm_014
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Title merge + styling (A1:I1 merged, 16pt bold white on teal)
  Component 2 (0.15) - Header row styling (row 3 bold with teal fill)
  Component 3 (0.15) - Summary row (A15:B15 merged, formulas in D15/E15/F15)
  Component 4 (0.15) - Data validations (Status and Priority dropdowns)
  Component 5 (0.15) - Conditional formatting (data bars on F, priority colors on H)
  Component 6 (0.20) - Line chart present with 2 series and correct title
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_014'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'ProjectDash' exists
    if 'ProjectDash' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ProjectDash' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ProjectDash']

    # ---------------------------------------------------------------
    # Component 1: Title merge + styling (0.20 points)
    # A1:I1 merged, 16pt bold, white font, teal (008080) fill
    # Initial has NO merges and plain styling -> this scores 0 on initial
    # ---------------------------------------------------------------
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        a1_merged = any('A1' in r and 'I1' in r for r in merged_ranges)

        cell_a1 = ws['A1']
        a1_bold = cell_a1.font.bold is True
        a1_size = cell_a1.font.size is not None and cell_a1.font.size >= 14

        # Check teal fill (FF008080)
        a1_fill_ok = False
        try:
            fg = cell_a1.fill.fgColor.rgb
            if fg and '008080' in str(fg):
                a1_fill_ok = True
        except Exception:
            pass

        # Check white font
        a1_font_white = False
        try:
            fc = cell_a1.font.color.rgb
            if fc and ('FFFFFF' in str(fc).upper()):
                a1_font_white = True
        except Exception:
            pass

        sub_score = 0.0
        if a1_merged:
            sub_score += 0.08
        if a1_bold and a1_size:
            sub_score += 0.04
        if a1_fill_ok:
            sub_score += 0.04
        if a1_font_white:
            sub_score += 0.04

        if sub_score > 0:
            print(f"PASS: Component 1 — Title styling (merged={a1_merged}, bold={a1_bold}, "
                  f"size={cell_a1.font.size}, fill_ok={a1_fill_ok}, white_font={a1_font_white}) "
                  f"({sub_score:.2f} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 1 — No title styling found (merged={a1_merged}, bold={a1_bold})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Header row styling (0.15 points)
    # Row 3 headers should be bold with teal fill
    # Initial has headers but no bold/fill
    # ---------------------------------------------------------------
    try:
        bold_count = 0
        fill_count = 0
        for col in range(1, 10):
            c = ws.cell(row=3, column=col)
            if c.font.bold is True:
                bold_count += 1
            try:
                fg = c.fill.fgColor.rgb
                if fg and '008080' in str(fg):
                    fill_count += 1
            except Exception:
                pass

        sub_score = 0.0
        if bold_count >= 7:
            sub_score += 0.075
        if fill_count >= 7:
            sub_score += 0.075

        if sub_score > 0:
            print(f"PASS: Component 2 — Header styling (bold={bold_count}/9, fill={fill_count}/9) "
                  f"({sub_score:.3f} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 2 — Headers not styled (bold={bold_count}/9, fill={fill_count}/9)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Summary row (0.15 points)
    # A15:B15 merged, D15=SUM(D4:D13), E15=SUM(E4:E13), F15=AVERAGE(F4:F13)
    # Initial has NO merge, NO formulas in D15/E15/F15
    # ---------------------------------------------------------------
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        a15_merged = any('A15' in r and 'B15' in r for r in merged_ranges)

        d15 = ws['D15'].value
        e15 = ws['E15'].value
        f15 = ws['F15'].value

        has_d15_formula = isinstance(d15, str) and 'SUM' in d15.upper() and 'D4' in d15.upper()
        has_e15_formula = isinstance(e15, str) and 'SUM' in e15.upper() and 'E4' in e15.upper()
        has_f15_formula = isinstance(f15, str) and 'AVERAGE' in f15.upper() and 'F4' in f15.upper()

        sub_score = 0.0
        if a15_merged:
            sub_score += 0.03
        if has_d15_formula:
            sub_score += 0.04
        if has_e15_formula:
            sub_score += 0.04
        if has_f15_formula:
            sub_score += 0.04

        if sub_score > 0:
            print(f"PASS: Component 3 — Summary row (merged={a15_merged}, D15={d15}, "
                  f"E15={e15}, F15={f15}) ({sub_score:.2f} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 3 — Summary row missing (merged={a15_merged}, "
                  f"D15={d15}, E15={e15}, F15={f15})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Data validations (0.15 points)
    # Status list on G4:G13, Priority list on H4:H13
    # Initial has NO data validations
    # ---------------------------------------------------------------
    try:
        dvs = ws.data_validations.dataValidation
        status_dv_found = False
        priority_dv_found = False

        for dv in dvs:
            if dv.type == 'list':
                formula = str(dv.formula1) if dv.formula1 else ''
                sqref = str(dv.sqref) if dv.sqref else ''

                # Status validation: contains 'To Do' or 'In Progress' or 'Done' or 'Blocked'
                if ('To Do' in formula or 'Done' in formula) and 'G' in sqref:
                    status_dv_found = True
                # Priority validation: contains 'High' or 'Medium' or 'Low'
                if ('High' in formula or 'Medium' in formula) and 'H' in sqref:
                    priority_dv_found = True

        sub_score = 0.0
        if status_dv_found:
            sub_score += 0.075
        if priority_dv_found:
            sub_score += 0.075

        if sub_score > 0:
            print(f"PASS: Component 4 — Data validations (status={status_dv_found}, "
                  f"priority={priority_dv_found}) ({sub_score:.3f} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 4 — No data validations found (count={len(dvs)})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Conditional formatting (0.15 points)
    # Data bars on F4:F13, priority color coding on H4:H13
    # Initial has NO conditional formatting
    # ---------------------------------------------------------------
    try:
        has_data_bar = False
        has_priority_cf = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar' and 'F' in cf_range:
                    has_data_bar = True
                if rule.type == 'cellIs' and 'H' in cf_range:
                    has_priority_cf = True

        sub_score = 0.0
        if has_data_bar:
            sub_score += 0.075
        if has_priority_cf:
            sub_score += 0.075

        if sub_score > 0:
            print(f"PASS: Component 5 — Conditional formatting (data_bar={has_data_bar}, "
                  f"priority_cf={has_priority_cf}) ({sub_score:.3f} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 5 — No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Line chart (0.20 points)
    # At least 1 chart, line type, 2 series, title containing 'Estimated' or 'Actual'
    # Initial has NO charts
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        has_chart = len(charts) >= 1

        sub_score = 0.0
        if has_chart:
            sub_score += 0.08
            chart = charts[0]

            # Check it's a line chart
            chart_class = chart.__class__.__name__
            if 'Line' in chart_class:
                sub_score += 0.04
                print(f"  Chart type: {chart_class} (line chart confirmed)")
            else:
                print(f"  Chart type: {chart_class} (expected LineChart)")

            # Check 2 series (estimated vs actual hours)
            if len(chart.series) >= 2:
                sub_score += 0.04
                print(f"  Chart series count: {len(chart.series)} (>= 2 confirmed)")
            else:
                print(f"  Chart series count: {len(chart.series)} (expected >= 2)")

            # Check chart title contains relevant keywords
            title_text = ''
            try:
                if chart.title and chart.title.tx and chart.title.tx.rich:
                    for p in chart.title.tx.rich.p:
                        for r in p.r:
                            if r.t:
                                title_text += r.t
            except Exception:
                pass

            if 'Estimated' in title_text or 'Actual' in title_text or 'Hours' in title_text:
                sub_score += 0.04
                print(f"  Chart title: '{title_text}' (relevant keywords found)")
            else:
                print(f"  Chart title: '{title_text}' (expected 'Hours: Estimated vs Actual')")

        if sub_score > 0:
            print(f"PASS: Component 6 — Chart ({sub_score:.2f} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 6 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
