"""
Reward Script: Apply thin top and bottom borders (no left/right) to A2:D10
Task ID: calc_fmt_border_top_bottom_only_078
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): All 36 cells in A2:D10 have thin top border
  Component 2 (0.5 pts): All 36 cells in A2:D10 have thin bottom border AND no left/right borders
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_border_top_bottom_only_078'
SHEET_NAME = 'Clean Table'


def verify_task(file_path):
    """
    Verify that thin top and bottom (but no left/right) borders have been
    applied to all cells in A2:D10.

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # --- Precondition: load workbook ---
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition: sheet must exist ---
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: All 36 cells in A2:D10 have thin TOP border (0.5 points)
    # This fails on initial file (no borders) and passes on golden file.
    try:
        top_pass_count = 0
        top_fail_cells = []
        for row in range(2, 11):       # rows 2-10
            for col in range(1, 5):    # columns A-D (1-4)
                cell = ws.cell(row=row, column=col)
                top_style = cell.border.top.style
                if top_style == 'thin':
                    top_pass_count += 1
                else:
                    from openpyxl.utils import get_column_letter
                    top_fail_cells.append(f"{get_column_letter(col)}{row}(top={top_style})")

        if top_pass_count == 36:
            print(f"PASS: Component 1 — all 36 cells in A2:D10 have thin top border (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — only {top_pass_count}/36 cells have thin top border")
            if top_fail_cells[:5]:
                print(f"  First failing cells: {top_fail_cells[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 (top border check) — {e}")

    # Component 2: All 36 cells in A2:D10 have thin BOTTOM border AND no LEFT or RIGHT border (0.5 points)
    # This fails on initial file (no borders) and passes on golden file.
    try:
        bottom_pass_count = 0
        no_side_pass_count = 0
        bottom_fail_cells = []
        side_fail_cells = []

        for row in range(2, 11):
            for col in range(1, 5):
                from openpyxl.utils import get_column_letter
                cell = ws.cell(row=row, column=col)
                coord = f"{get_column_letter(col)}{row}"
                b = cell.border
                if b.bottom.style == 'thin':
                    bottom_pass_count += 1
                else:
                    bottom_fail_cells.append(f"{coord}(bottom={b.bottom.style})")
                if b.left.style is None and b.right.style is None:
                    no_side_pass_count += 1
                else:
                    side_fail_cells.append(f"{coord}(left={b.left.style},right={b.right.style})")

        if bottom_pass_count == 36 and no_side_pass_count == 36:
            print(f"PASS: Component 2 — all 36 cells have thin bottom border and no left/right borders (0.5 pts)")
            total_score += 0.5
        else:
            if bottom_pass_count < 36:
                print(f"FAIL: Component 2 — only {bottom_pass_count}/36 cells have thin bottom border")
                if bottom_fail_cells[:5]:
                    print(f"  First failing cells: {bottom_fail_cells[:5]}")
            if no_side_pass_count < 36:
                print(f"FAIL: Component 2 — {36 - no_side_pass_count} cells have unwanted left/right borders")
                if side_fail_cells[:5]:
                    print(f"  First offending cells: {side_fail_cells[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 (bottom/sides border check) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
