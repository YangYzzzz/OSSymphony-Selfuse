"""
Initial Setup: Apply thin top and bottom borders to A2:D10
Task ID: calc_fmt_border_top_bottom_only_078
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_border_top_bottom_only_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Clean Table'

    # --- Header row ---
    headers = ['Name', 'Value', 'Date', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Data rows 2-10 (9 rows, realistic content) ---
    data = [
        ['Sarah Chen',      45230.50,  '2025-01-15', 'Q1 budget approved'],
        ['Marcus Johnson',  32800.00,  '2025-01-22', 'Pending review'],
        ['Priya Nair',      58640.75,  '2025-02-03', 'Contract renewed'],
        ['David Okafor',    27150.00,  '2025-02-14', 'Onboarding complete'],
        ['Emma Lindqvist',  61200.00,  '2025-02-28', 'Annual audit done'],
        ['Carlos Rivera',   39500.25,  '2025-03-05', 'Invoice sent'],
        ['Yuki Tanaka',     47800.00,  '2025-03-12', 'Awaiting sign-off'],
        ['Aisha Patel',     53100.50,  '2025-03-20', 'Phase 2 initiated'],
        ['Tom Whitfield',   29900.00,  '2025-03-28', 'Follow-up required'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 24

    # NOTE: NO borders are applied — the task requires adding them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
