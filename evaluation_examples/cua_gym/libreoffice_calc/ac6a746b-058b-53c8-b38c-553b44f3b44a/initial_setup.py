"""
Initial Setup: Create QC spreadsheet with production line defect data
Task ID: calc_ops_023
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: QC ---
    ws = wb.active
    ws.title = 'QC'

    # Headers
    headers = ['Production Line', 'Units Produced', 'Defects', 'Defect Rate']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Data rows
    data = [
        ['Line 1', 10000, 45, 0.0045],
        ['Line 2', 8500, 310, 0.0365],
        ['Line 3', 12000, 96, 0.008],
        ['Line 4', 9000, 180, 0.02],
        ['Line 5', 11000, 55, 0.005],
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = data_border
            if c == 1:
                cell.font = Font(name='Calibri', size=11)
            elif c in (2, 3):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal="right")
            elif c == 4:
                cell.number_format = '0.00%'
                cell.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
