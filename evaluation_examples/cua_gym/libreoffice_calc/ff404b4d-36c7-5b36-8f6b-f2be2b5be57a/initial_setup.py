"""
Initial Setup: Retail Markdown Optimization Sheet
Task ID: calc_wf_075
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Markdown"

    # --- Headers ---
    headers = [
        "SKU", "Product Name", "Original Price", "Current Price", "Unit Cost",
        "Inventory", "Weekly Sales Rate", "Target Weeks to Clear"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- 20 SKUs of realistic retail data ---
    data = [
        ["SKU-1001", "Classic Denim Jacket", 89.99, 89.99, 38.50, 245, 12, 8],
        ["SKU-1002", "Leather Crossbody Bag", 129.00, 129.00, 52.00, 180, 8, 10],
        ["SKU-1003", "Running Shoes Pro X", 159.95, 159.95, 65.00, 320, 15, 12],
        ["SKU-1004", "Cashmere Blend Sweater", 198.00, 198.00, 78.00, 95, 4, 6],
        ["SKU-1005", "Wireless Earbuds Elite", 79.99, 79.99, 28.00, 410, 22, 10],
        ["SKU-1006", "Yoga Mat Premium", 54.95, 54.95, 18.50, 175, 10, 8],
        ["SKU-1007", "Stainless Steel Watch", 245.00, 245.00, 95.00, 60, 3, 10],
        ["SKU-1008", "Organic Cotton T-Shirt", 34.99, 34.99, 12.00, 520, 30, 12],
        ["SKU-1009", "Smart Home Hub", 149.99, 149.99, 62.00, 88, 5, 8],
        ["SKU-1010", "Hiking Backpack 40L", 119.00, 119.00, 45.00, 150, 7, 10],
        ["SKU-1011", "Polarized Sunglasses", 65.00, 65.00, 22.00, 290, 18, 8],
        ["SKU-1012", "Chef's Knife Set", 189.99, 189.99, 72.00, 45, 2, 12],
        ["SKU-1013", "Bluetooth Speaker Mini", 44.99, 44.99, 15.00, 380, 25, 6],
        ["SKU-1014", "Winter Parka Insulated", 279.00, 279.00, 110.00, 72, 3, 8],
        ["SKU-1015", "Ceramic Plant Pot Set", 42.50, 42.50, 14.00, 200, 12, 10],
        ["SKU-1016", "Fitness Tracker Band", 99.95, 99.95, 35.00, 165, 9, 8],
        ["SKU-1017", "Silk Scarf Collection", 78.00, 78.00, 28.00, 110, 5, 10],
        ["SKU-1018", "Espresso Machine Pro", 349.99, 349.99, 145.00, 35, 2, 8],
        ["SKU-1019", "Canvas Tote Bag", 28.99, 28.99, 9.50, 450, 28, 10],
        ["SKU-1020", "LED Desk Lamp Smart", 67.50, 67.50, 24.00, 130, 8, 6],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (3, 4, 5):  # price columns
                cell.number_format = '$#,##0.00'
            elif c == 6:  # inventory
                cell.number_format = '#,##0'
            elif c in (7, 8):  # rates/weeks
                cell.number_format = '0'
            if c >= 3:
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = {"A": 12, "B": 26, "C": 15, "D": 14, "E": 12, "F": 12, "G": 18, "H": 22}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    ws.sheet_view.zoomScale = 100

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
