"""
Reward Script: Retail Markdown Optimization Sheet
Task ID: calc_wf_075
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Weeks of Supply & Acceleration Factor columns (I, J) with formulas
  Component 2 (0.25) - Revenue projection columns (K-O) at 5 discount levels with elasticity
  Component 3 (0.20) - Margin columns (P-T) and Optimal Discount column (U)
  Component 4 (0.15) - Chart comparing revenue vs margin by discount level
  Component 5 (0.10) - Summary area with averages for chart data (rows 28-33)
  Component 6 (0.10) - Conditional formatting on WOS column (>12 weeks = red)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_075'


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        import time
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Markdown' sheet must exist
    if 'Markdown' not in wb.sheetnames:
        print("FAIL: 'Markdown' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Markdown']

    # =========================================================================
    # Component 1: Weeks of Supply (col I) and Required Acceleration Factor (col J)
    # (0.20 points)
    # Initial env has NO columns I or J. Golden has formulas =F/G and =I/H.
    # =========================================================================
    try:
        wos_formulas_ok = 0
        accel_formulas_ok = 0
        for row in range(2, 22):  # rows 2-21 (20 SKUs)
            val_i = ws.cell(row=row, column=9).value  # col I
            val_j = ws.cell(row=row, column=10).value  # col J
            if val_i is not None:
                val_str = str(val_i).upper().replace(" ", "")
                # Check for WOS formula: =F/G pattern
                if "F" in val_str and "G" in val_str and ("/" in val_str or "F" in val_str):
                    wos_formulas_ok += 1
                elif isinstance(val_i, (int, float)):
                    # Also accept computed values (if file was saved by Calc)
                    wos_formulas_ok += 1
            if val_j is not None:
                val_str_j = str(val_j).upper().replace(" ", "")
                if ("I" in val_str_j or "F" in val_str_j) and ("H" in val_str_j or "/" in val_str_j):
                    accel_formulas_ok += 1
                elif isinstance(val_j, (int, float)):
                    accel_formulas_ok += 1

        # Header check for col I
        header_i = ws.cell(row=1, column=9).value
        header_j = ws.cell(row=1, column=10).value
        headers_present = (header_i is not None and header_j is not None)

        if wos_formulas_ok >= 15 and accel_formulas_ok >= 15 and headers_present:
            print(f"PASS: Component 1 — WOS formulas: {wos_formulas_ok}/20, Accel formulas: {accel_formulas_ok}/20, headers present (0.20 pts)")
            total_score += 0.20
        elif wos_formulas_ok >= 10 or accel_formulas_ok >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 1 — WOS: {wos_formulas_ok}/20, Accel: {accel_formulas_ok}/20 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — WOS formulas: {wos_formulas_ok}/20, Accel formulas: {accel_formulas_ok}/20")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Revenue projection columns K-O at 5 discount levels (0.25 pts)
    # Formula pattern: =G*multiplier*H*C*(1-discount)
    # Elasticity: each 10% discount increases velocity by 25%
    #   10% -> 1.25x, 20% -> 1.5x, 30% -> 1.75x, 40% -> 2.0x, 50% -> 2.25x
    # =========================================================================
    try:
        rev_cols_ok = 0
        # Check columns K(11) through O(15)
        for col_idx in range(11, 16):
            col_has_data = 0
            for row in range(2, 22):
                val = ws.cell(row=row, column=col_idx).value
                if val is not None:
                    val_str = str(val).upper().replace(" ", "")
                    # Should reference G (velocity), C (price), and contain multiplication
                    if isinstance(val, str) and ("G" in val_str and "C" in val_str):
                        col_has_data += 1
                    elif isinstance(val, (int, float)) and val > 0:
                        col_has_data += 1
            if col_has_data >= 15:
                rev_cols_ok += 1

        # Also check headers exist for these columns
        rev_headers = sum(1 for c in range(11, 16) if ws.cell(row=1, column=c).value is not None)

        if rev_cols_ok >= 4 and rev_headers >= 4:
            print(f"PASS: Component 2 — {rev_cols_ok}/5 revenue columns populated with formulas, {rev_headers}/5 headers (0.25 pts)")
            total_score += 0.25
        elif rev_cols_ok >= 2:
            partial = 0.12
            print(f"PARTIAL: Component 2 — {rev_cols_ok}/5 revenue columns ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {rev_cols_ok}/5 revenue columns populated")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Margin columns (P-T) and Optimal Discount column (U) (0.20 pts)
    # Margin = Revenue - Cost (P=K-cost, Q=L-cost, etc.)
    # Optimal = IF(MAX(P:T)=P,"10%",...) nested IF
    # =========================================================================
    try:
        margin_cols_ok = 0
        # Check columns P(16) through T(20)
        for col_idx in range(16, 21):
            col_has_data = 0
            for row in range(2, 22):
                val = ws.cell(row=row, column=col_idx).value
                if val is not None:
                    val_str = str(val).upper().replace(" ", "")
                    if isinstance(val, str) and ("E" in val_str or "-" in val_str):
                        col_has_data += 1
                    elif isinstance(val, (int, float)):
                        col_has_data += 1
            if col_has_data >= 15:
                margin_cols_ok += 1

        # Check column U (21) — Optimal Discount
        optimal_ok = 0
        for row in range(2, 22):
            val_u = ws.cell(row=row, column=21).value
            if val_u is not None:
                val_str_u = str(val_u).upper().replace(" ", "")
                if isinstance(val_u, str) and ("MAX" in val_str_u or "IF" in val_str_u):
                    optimal_ok += 1
                elif isinstance(val_u, str) and "%" in val_str_u:
                    optimal_ok += 1

        margin_headers = sum(1 for c in range(16, 22) if ws.cell(row=1, column=c).value is not None)
        optimal_header = ws.cell(row=1, column=21).value is not None

        if margin_cols_ok >= 4 and optimal_ok >= 15 and optimal_header:
            print(f"PASS: Component 3 — {margin_cols_ok}/5 margin cols, {optimal_ok}/20 optimal formulas, headers: {margin_headers} (0.20 pts)")
            total_score += 0.20
        elif margin_cols_ok >= 2 or optimal_ok >= 10:
            partial = 0.10
            print(f"PARTIAL: Component 3 — margin: {margin_cols_ok}/5, optimal: {optimal_ok}/20 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — margin cols: {margin_cols_ok}/5, optimal: {optimal_ok}/20")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Chart comparing Revenue vs Margin at each discount level (0.15 pts)
    # Expected: BarChart with 2 series, title mentioning revenue/margin/discount
    # =========================================================================
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            has_two_series = len(chart.series) >= 2

            # Extract chart title text
            chart_title_text = ""
            if chart.title and hasattr(chart.title, 'tx') and chart.title.tx:
                if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                    for p in chart.title.tx.rich.p:
                        for r in p.r:
                            if r.t:
                                chart_title_text += r.t

            if has_two_series:
                print(f"PASS: Component 4 — Chart found with {len(chart.series)} series, title='{chart_title_text}' (0.15 pts)")
                total_score += 0.15
            else:
                print(f"PARTIAL: Component 4 — Chart found but only {len(chart.series)} series (need >=2)")
                total_score += 0.07
        else:
            print(f"FAIL: Component 4 — No charts found in 'Markdown' sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Summary area with averages for chart data (rows 28-33) (0.10 pts)
    # Expected: Discount Level labels (10%-50%) in col A, AVERAGE formulas in B and C
    # =========================================================================
    try:
        summary_rows_ok = 0
        # Check for discount level labels and average formulas in rows 28-33 area
        # Search a range since exact row placement may vary
        found_labels = 0
        found_avg_formulas = 0
        for row in range(23, 40):
            val_a = ws.cell(row=row, column=1).value
            val_b = ws.cell(row=row, column=2).value
            val_c = ws.cell(row=row, column=3).value
            if val_a is not None and "%" in str(val_a):
                found_labels += 1
                if val_b is not None:
                    b_str = str(val_b).upper().replace(" ", "")
                    if "AVERAGE" in b_str or isinstance(val_b, (int, float)):
                        found_avg_formulas += 1
                if val_c is not None:
                    c_str = str(val_c).upper().replace(" ", "")
                    if "AVERAGE" in c_str or isinstance(val_c, (int, float)):
                        found_avg_formulas += 1

        if found_labels >= 4 and found_avg_formulas >= 6:
            print(f"PASS: Component 5 — Summary area: {found_labels} discount labels, {found_avg_formulas} avg formulas (0.10 pts)")
            total_score += 0.10
        elif found_labels >= 2 or found_avg_formulas >= 2:
            partial = 0.05
            print(f"PARTIAL: Component 5 — labels: {found_labels}, avg formulas: {found_avg_formulas} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — labels: {found_labels}, avg formulas: {found_avg_formulas}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Conditional formatting on WOS column (>12 weeks = red) (0.10 pts)
    # Expected: CellIsRule on I2:I21, operator=greaterThan, formula=['12'], red fill
    # =========================================================================
    try:
        cf_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the range covers column I (WOS column)
            if "I" in cf_range:
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', '')
                    rule_op = getattr(rule, 'operator', '')
                    rule_formula = getattr(rule, 'formula', [])
                    # Check for "greater than 12" rule
                    if rule_type == 'cellIs' and rule_op == 'greaterThan':
                        if rule_formula and '12' in str(rule_formula[0]):
                            cf_found = True
                            # Verify red fill
                            has_red = False
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                try:
                                    fg_rgb = rule.dxf.fill.fgColor.rgb
                                    if fg_rgb and 'FF0000' in str(fg_rgb):
                                        has_red = True
                                except:
                                    pass
                            if has_red:
                                print(f"PASS: Component 6 — Conditional formatting on {cf_range}: >12 with red fill (0.10 pts)")
                                total_score += 0.10
                            else:
                                print(f"PARTIAL: Component 6 — CF rule >12 found but fill not confirmed red (0.05 pts)")
                                total_score += 0.05

        if not cf_found:
            print(f"FAIL: Component 6 — No conditional formatting with >12 rule found on WOS column")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
