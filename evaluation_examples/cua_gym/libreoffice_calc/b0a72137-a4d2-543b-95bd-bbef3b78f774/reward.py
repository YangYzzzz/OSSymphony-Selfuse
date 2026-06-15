"""
Reward Script: Build a workout log with exercise tracking, personal records highlighting, and volume chart.
Task ID: calc_gpm_054
Domain: libreoffice_calc
Scoring:
  Component 1: Volume formulas in F4:F18 (0.20)
  Component 2: 1RM estimate formulas in G4:G18 (0.15)
  Component 3: PR detection formulas in H4:H18 (0.15)
  Component 4: Exercise Summary section rows 20-25 with MAXIFS/SUMIFS (0.20)
  Component 5: Bar chart present with correct title (0.15)
  Component 6: Conditional formatting rules (H, F, E columns) (0.10)
  Component 7: Number formats #,##0 on E-G columns (0.05)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_054'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print("CRITICAL: Cannot load file %s: %s" % (file_path, e))
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Workout' not in wb.sheetnames:
        print("CRITICAL: 'Workout' sheet not found. Sheets: %s" % wb.sheetnames)
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Workout']

    # Component 1: Volume formulas in F4:F18 (0.20 points)
    # These should be =C*D*E style formulas. Initial has None in these cells.
    try:
        volume_count = 0
        for r in range(4, 19):
            val = ws.cell(row=r, column=6).value  # column F
            if val is not None and isinstance(val, str) and '=' in val:
                # Check it references C, D, E columns (multiplication of sets*reps*weight)
                v_upper = val.upper().replace(' ', '')
                if ('C%d' % r in v_upper or 'C$' in v_upper) and ('D%d' % r in v_upper or 'D$' in v_upper) and ('E%d' % r in v_upper or 'E$' in v_upper):
                    volume_count += 1
                elif '*' in v_upper:
                    # Accept any multiplicative formula referencing the row
                    volume_count += 1
        if volume_count >= 13:  # at least 13 out of 15
            print("PASS: Component 1 - Volume formulas found in %d/15 cells (0.20 pts)" % volume_count)
            total_score += 0.20
        elif volume_count >= 8:
            partial = 0.10
            print("PARTIAL: Component 1 - Volume formulas found in %d/15 cells (%.2f pts)" % (volume_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 1 - Volume formulas found in only %d/15 cells" % volume_count)
    except Exception as e:
        print("ERROR: Component 1 - %s" % e)

    # Component 2: 1RM estimate formulas in G4:G18 (0.15 points)
    # Epley formula: =E*(1+D/30)
    try:
        rm_count = 0
        for r in range(4, 19):
            val = ws.cell(row=r, column=7).value  # column G
            if val is not None and isinstance(val, str) and '=' in val:
                v_upper = val.upper().replace(' ', '')
                # Check references to E and D columns (Epley formula pattern)
                if ('E%d' % r in v_upper or 'E$' in v_upper) and ('D%d' % r in v_upper or 'D$' in v_upper):
                    rm_count += 1
                elif '30' in v_upper:
                    # Epley formula contains /30
                    rm_count += 1
        if rm_count >= 13:
            print("PASS: Component 2 - 1RM formulas found in %d/15 cells (0.15 pts)" % rm_count)
            total_score += 0.15
        elif rm_count >= 8:
            partial = 0.08
            print("PARTIAL: Component 2 - 1RM formulas found in %d/15 cells (%.2f pts)" % (rm_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 2 - 1RM formulas found in only %d/15 cells" % rm_count)
    except Exception as e:
        print("ERROR: Component 2 - %s" % e)

    # Component 3: PR detection formulas in H4:H18 (0.15 points)
    # =IF(G=MAX/MAXIFS of same exercise, "NEW PR!", "")
    try:
        pr_count = 0
        for r in range(4, 19):
            val = ws.cell(row=r, column=8).value  # column H
            if val is not None and isinstance(val, str) and '=' in val:
                v_upper = val.upper().replace(' ', '')
                # Check for IF with MAX/MAXIFS and "NEW PR!" or "NEWPR!"
                if 'IF(' in v_upper and ('MAX' in v_upper) and ('PR' in v_upper or 'pr' in val):
                    pr_count += 1
        if pr_count >= 13:
            print("PASS: Component 3 - PR detection formulas found in %d/15 cells (0.15 pts)" % pr_count)
            total_score += 0.15
        elif pr_count >= 8:
            partial = 0.08
            print("PARTIAL: Component 3 - PR detection formulas found in %d/15 cells (%.2f pts)" % (pr_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 3 - PR detection formulas found in only %d/15 cells" % pr_count)
    except Exception as e:
        print("ERROR: Component 3 - %s" % e)

    # Component 4: Exercise Summary section (0.20 points)
    # Row 20: merged A20:B20 with "Exercise Summary", C20: "Best 1RM", D20: "Total Volume"
    # Rows 21-25: exercise names + MAXIFS + SUMIFS formulas
    try:
        summary_score = 0.0

        # Check A20 has "Exercise Summary"
        a20_val = ws['A20'].value
        if a20_val is not None and 'exercise' in str(a20_val).lower() and 'summary' in str(a20_val).lower():
            summary_score += 0.04
            print("  Sub-check: A20 has Exercise Summary text")
        else:
            print("  Sub-fail: A20 value is %r, expected 'Exercise Summary'" % a20_val)

        # Check A20:B20 merged
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        a20_merged = any('A20' in mr and 'B20' in mr for mr in merged_ranges)
        if a20_merged:
            summary_score += 0.02
            print("  Sub-check: A20:B20 is merged")
        else:
            print("  Sub-fail: A20:B20 not merged. Ranges: %s" % merged_ranges)

        # Check C20 and D20 headers
        c20 = ws['C20'].value
        d20 = ws['D20'].value
        if c20 is not None and '1rm' in str(c20).lower():
            summary_score += 0.02
            print("  Sub-check: C20 has 1RM header")
        if d20 is not None and 'volume' in str(d20).lower():
            summary_score += 0.02
            print("  Sub-check: D20 has Volume header")

        # Check rows 21-25 have exercise names and MAXIFS/SUMIFS
        exercises_found = 0
        maxifs_found = 0
        sumifs_found = 0
        expected_exercises = ['bench press', 'squat', 'deadlift', 'ohp', 'row']
        for r in range(21, 26):
            a_val = ws.cell(row=r, column=1).value
            c_val = ws.cell(row=r, column=3).value
            d_val = ws.cell(row=r, column=4).value

            if a_val is not None and str(a_val).lower().strip() in expected_exercises:
                exercises_found += 1

            if c_val is not None and isinstance(c_val, str) and 'MAXIFS' in c_val.upper():
                maxifs_found += 1

            if d_val is not None and isinstance(d_val, str) and 'SUMIFS' in d_val.upper():
                sumifs_found += 1

        if exercises_found >= 4:
            summary_score += 0.03
            print("  Sub-check: %d/5 exercise names found" % exercises_found)
        else:
            print("  Sub-fail: Only %d/5 exercise names found" % exercises_found)

        if maxifs_found >= 4:
            summary_score += 0.04
            print("  Sub-check: %d/5 MAXIFS formulas found" % maxifs_found)
        else:
            print("  Sub-fail: Only %d/5 MAXIFS formulas found" % maxifs_found)

        if sumifs_found >= 4:
            summary_score += 0.03
            print("  Sub-check: %d/5 SUMIFS formulas found" % sumifs_found)
        else:
            print("  Sub-fail: Only %d/5 SUMIFS formulas found" % sumifs_found)

        if summary_score > 0:
            print("PASS: Component 4 - Exercise Summary (%.2f pts)" % summary_score)
        else:
            print("FAIL: Component 4 - Exercise Summary section missing")
        total_score += summary_score
    except Exception as e:
        print("ERROR: Component 4 - %s" % e)

    # Component 5: Bar chart present with title (0.15 points)
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            # Check chart title contains "volume" or "training"
            chart_title_text = ''
            if chart.title is not None:
                try:
                    # Extract text from title object
                    if hasattr(chart.title, 'tx') and chart.title.tx is not None:
                        if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich is not None:
                            for p in chart.title.tx.rich.p:
                                for r in p.r:
                                    if r.t:
                                        chart_title_text += r.t
                except Exception:
                    pass

            if 'volume' in chart_title_text.lower():
                print("PASS: Component 5 - Chart found with title '%s' (0.15 pts)" % chart_title_text)
                total_score += 0.15
            elif len(charts) >= 1:
                # Chart exists but title may not match exactly
                print("PARTIAL: Component 5 - Chart found but title='%s' (0.10 pts)" % chart_title_text)
                total_score += 0.10
        else:
            print("FAIL: Component 5 - No charts found")
    except Exception as e:
        print("ERROR: Component 5 - %s" % e)

    # Component 6: Conditional formatting rules (0.10 points)
    # Expected: H4:H18 cellIs "NEW PR!", F4:F18 dataBar, E4:E18 expression/top3
    try:
        cf_rules = list(ws.conditional_formatting)
        cf_count = len(cf_rules)

        cf_score = 0.0
        has_pr_cf = False
        has_databar_cf = False
        has_weight_cf = False

        for cf in cf_rules:
            cf_range = str(cf)
            for rule in cf.rules:
                rtype = rule.type
                # PR highlighting on H column
                if 'H' in cf_range and rtype == 'cellIs':
                    has_pr_cf = True
                # Data bars on F column
                if 'F' in cf_range and rtype == 'dataBar':
                    has_databar_cf = True
                # Top 3 on E column
                if 'E' in cf_range and (rtype == 'expression' or rtype == 'top10'):
                    has_weight_cf = True

        if has_pr_cf:
            cf_score += 0.04
            print("  Sub-check: PR conditional formatting found on H column")
        else:
            print("  Sub-fail: No PR conditional formatting on H column")

        if has_databar_cf:
            cf_score += 0.03
            print("  Sub-check: Data bars found on F column")
        else:
            print("  Sub-fail: No data bars on F column")

        if has_weight_cf:
            cf_score += 0.03
            print("  Sub-check: Weight conditional formatting found on E column")
        else:
            print("  Sub-fail: No weight conditional formatting on E column")

        if cf_score > 0:
            print("PASS: Component 6 - Conditional formatting (%.2f pts)" % cf_score)
        else:
            print("FAIL: Component 6 - No conditional formatting rules found")
        total_score += cf_score
    except Exception as e:
        print("ERROR: Component 6 - %s" % e)

    # Component 7: Number formats #,##0 on E-G columns (0.05 points)
    try:
        nf_count = 0
        for col in [5, 6, 7]:  # E, F, G
            nf = ws.cell(row=4, column=col).number_format
            if nf is not None and '#' in str(nf) and '0' in str(nf):
                nf_count += 1
        if nf_count >= 3:
            print("PASS: Component 7 - Number format #,##0 on E-G columns (0.05 pts)")
            total_score += 0.05
        elif nf_count >= 1:
            partial = 0.02
            print("PARTIAL: Component 7 - Number format on %d/3 columns (%.2f pts)" % (nf_count, partial))
            total_score += partial
        else:
            print("FAIL: Component 7 - Number format is General on E-G columns")
    except Exception as e:
        print("ERROR: Component 7 - %s" % e)

    final_score = round(min(total_score, 1.0), 2)
    print("\nScore: %.2f/1.0" % total_score)
    print("REWARD: %.1f" % final_score)
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print("PERSIST_WARN: save hook failed: %s" % e)


# Entry point
file_path = '%s/%s.xlsx' % (WORKDIR, TASK_ID)
if not os.path.exists(file_path):
    print("File not found: %s" % file_path)
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
