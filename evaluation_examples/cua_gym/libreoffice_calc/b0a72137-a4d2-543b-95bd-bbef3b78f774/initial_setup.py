"""
Initial Setup: Build a workout log with exercise tracking data (pre-task state)
Task ID: calc_gpm_054
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Workout'

    # --- Title Row: Merge A1:H1 ---
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'Strength Training Log - March 2026'
    title_cell.font = Font(size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='FF8B0000', end_color='FF8B0000', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 3: Headers ---
    headers = ['Date', 'Exercise', 'Sets', 'Reps', 'Weight (lbs)', 'Volume', '1RM Est', 'PR?']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data Rows 4-18: 15 workout entries ---
    workout_data = [
        (date(2026, 3, 2),  'Bench Press', 4, 8,  185),
        (date(2026, 3, 2),  'Squat',       5, 5,  275),
        (date(2026, 3, 4),  'Deadlift',    3, 5,  315),
        (date(2026, 3, 4),  'OHP',         4, 8,  115),
        (date(2026, 3, 6),  'Row',         4, 10, 155),
        (date(2026, 3, 9),  'Bench Press', 5, 5,  205),
        (date(2026, 3, 9),  'Squat',       4, 8,  245),
        (date(2026, 3, 11), 'Deadlift',    5, 3,  365),
        (date(2026, 3, 11), 'OHP',         3, 10, 105),
        (date(2026, 3, 13), 'Row',         5, 8,  165),
        (date(2026, 3, 16), 'Bench Press', 4, 6,  195),
        (date(2026, 3, 16), 'Squat',       5, 5,  285),
        (date(2026, 3, 18), 'Deadlift',    4, 5,  335),
        (date(2026, 3, 20), 'OHP',         4, 6,  120),
        (date(2026, 3, 20), 'Row',         4, 8,  170),
    ]

    for r, (dt, exercise, sets, reps, weight) in enumerate(workout_data, 4):
        ws.cell(row=r, column=1, value=dt)
        ws.cell(row=r, column=1).number_format = 'MMM DD'
        ws.cell(row=r, column=2, value=exercise)
        ws.cell(row=r, column=3, value=sets)
        ws.cell(row=r, column=4, value=reps)
        ws.cell(row=r, column=5, value=weight)
        # F (Volume), G (1RM Est), H (PR?) left EMPTY intentionally

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
