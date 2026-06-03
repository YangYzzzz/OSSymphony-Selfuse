"""
Initial Setup: COUNTIFS multi-criteria formula task
Task ID: calc_fmb_countifs_multi_010
Domain: libreoffice_calc
"""

import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_countifs_multi_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HR Records'

    # --- Headers (Row 1) ---
    headers = ['Employee ID', 'Name', 'Department', 'Salary', 'Years', 'Rating']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Department distribution:
    # Engineering: 95 employees (38 with salary > 90000, 57 with salary <= 90000)
    # Marketing: 60
    # Sales: 75
    # Finance: 45
    # Operations: 25
    # Total: 300

    first_names = [
        'James', 'Mary', 'John', 'Patricia', 'Robert', 'Jennifer', 'Michael',
        'Linda', 'William', 'Barbara', 'David', 'Elizabeth', 'Richard', 'Susan',
        'Joseph', 'Jessica', 'Thomas', 'Sarah', 'Charles', 'Karen', 'Christopher',
        'Lisa', 'Daniel', 'Nancy', 'Matthew', 'Betty', 'Anthony', 'Margaret',
        'Mark', 'Sandra', 'Donald', 'Ashley', 'Steven', 'Kimberly', 'Paul',
        'Emily', 'Andrew', 'Donna', 'Joshua', 'Michelle', 'Kenneth', 'Dorothy',
        'Kevin', 'Carol', 'Brian', 'Amanda', 'George', 'Melissa', 'Timothy', 'Deborah',
        'Ronald', 'Stephanie', 'Edward', 'Rebecca', 'Jason', 'Sharon', 'Jeffrey',
        'Laura', 'Ryan', 'Cynthia', 'Jacob', 'Kathleen', 'Gary', 'Amy', 'Nicholas',
        'Angela', 'Eric', 'Shirley', 'Jonathan', 'Anna', 'Stephen', 'Brenda',
        'Larry', 'Pamela', 'Justin', 'Emma', 'Scott', 'Nicole', 'Brandon', 'Helen',
        'Benjamin', 'Samantha', 'Samuel', 'Katherine', 'Frank', 'Christine',
        'Gregory', 'Debra', 'Alexander', 'Rachel', 'Patrick', 'Carolyn',
        'Raymond', 'Janet', 'Jack', 'Catherine', 'Dennis', 'Maria', 'Jerry'
    ]

    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
        'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
        'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
        'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
        'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
        'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green',
        'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts', 'Chen', 'Patel', 'Kim', 'Park', 'Gupta', 'Singh',
        'Kumar', 'Shah', 'Reed', 'Collins', 'Stewart', 'Morris', 'Morales',
        'Murphy', 'Cook', 'Rogers', 'Gutierrez', 'Ortiz', 'Morgan', 'Cooper',
        'Peterson', 'Bailey', 'Reed', 'Kelly', 'Howard', 'Ramos', 'Cox',
        'Ward', 'Richardson', 'Watson', 'Brooks', 'Chavez', 'Wood', 'James',
        'Bennett', 'Gray', 'Mendoza', 'Ruiz', 'Hughes', 'Price', 'Alvarez',
        'Castillo', 'Sanders', 'Patel', 'Myers', 'Long', 'Ross', 'Foster'
    ]

    # Build records list
    # Engineering: 38 with salary > 90000, 57 with salary <= 90000
    # Marketing: 60 (mix of salaries)
    # Sales: 75 (mix of salaries)
    # Finance: 45 (mix of salaries)
    # Operations: 25 (mix of salaries)

    records = []
    emp_id = 1001

    def rand_name():
        return f"{random.choice(first_names)} {random.choice(last_names)}"

    def rand_years():
        return random.randint(1, 18)

    def rand_rating():
        return round(random.uniform(2.5, 5.0), 1)

    # Engineering: 38 employees with salary > 90000
    for i in range(38):
        salary = random.randint(91000, 148000)
        records.append([f'EMP{emp_id:04d}', rand_name(), 'Engineering', salary, rand_years(), rand_rating()])
        emp_id += 1

    # Engineering: 57 employees with salary <= 90000
    for i in range(57):
        salary = random.randint(55000, 90000)
        records.append([f'EMP{emp_id:04d}', rand_name(), 'Engineering', salary, rand_years(), rand_rating()])
        emp_id += 1

    # Marketing: 60 employees
    for i in range(60):
        salary = random.randint(48000, 125000)
        records.append([f'EMP{emp_id:04d}', rand_name(), 'Marketing', salary, rand_years(), rand_rating()])
        emp_id += 1

    # Sales: 75 employees
    for i in range(75):
        salary = random.randint(42000, 115000)
        records.append([f'EMP{emp_id:04d}', rand_name(), 'Sales', salary, rand_years(), rand_rating()])
        emp_id += 1

    # Finance: 45 employees
    for i in range(45):
        salary = random.randint(52000, 130000)
        records.append([f'EMP{emp_id:04d}', rand_name(), 'Finance', salary, rand_years(), rand_rating()])
        emp_id += 1

    # Operations: 25 employees
    for i in range(25):
        salary = random.randint(45000, 98000)
        records.append([f'EMP{emp_id:04d}', rand_name(), 'Operations', salary, rand_years(), rand_rating()])
        emp_id += 1

    # Shuffle so departments are intermixed
    random.shuffle(records)

    # Renumber Employee IDs after shuffle
    for i, rec in enumerate(records):
        rec[0] = f'EMP{1001 + i:04d}'

    # Write records to sheet (rows 2-301)
    for r, rec in enumerate(records, 2):
        for c, val in enumerate(rec, 1):
            ws.cell(row=r, column=c, value=val)

    # F3: label for the target formula
    ws['F3'] = 'Eng >$90K'
    # G3: EMPTY — this is the target cell for the COUNTIFS formula

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verification
    wb2 = openpyxl.load_workbook(OUTPUT)
    ws2 = wb2['HR Records']
    eng_high = sum(
        1 for r in range(2, 302)
        if ws2.cell(row=r, column=3).value == 'Engineering'
        and isinstance(ws2.cell(row=r, column=4).value, (int, float))
        and ws2.cell(row=r, column=4).value > 90000
    )
    print(f'Verification: Engineering employees with salary > $90,000: {eng_high}')
    print(f'G3 value (should be empty): {ws2["G3"].value}')
    print(f'F3 value (should be "Eng >$90K"): {ws2["F3"].value}')

create_initial()
