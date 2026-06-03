"""
Reward Script: Count Engineering employees with salary >$90,000 using COUNTIFS
Task ID: calc_fmb_countifs_multi_010
Domain: libreoffice_calc
Scoring:
  Component 1: G3 contains any formula (0.4 pts)
  Component 2: G3 formula is COUNTIFS with correct ranges and criteria (0.6 pts)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_countifs_multi_010'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    Task: Put =COUNTIFS(C2:C301,"Engineering",D2:D301,">90000") in cell G3
    of the 'HR Records' sheet. The formula should count Engineering employees
    with salary above $90,000 (expected result: 38).

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must load successfully
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'HR Records' sheet must exist
    if 'HR Records' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'HR Records' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['HR Records']

    # Component 1: G3 contains a formula (0.4 points)
    # Initial state: G3 is None (empty)
    # Golden state: G3 has a COUNTIFS formula
    # This FAILS on initial (G3 is empty) and PASSES on golden
    try:
        g3_value = ws['G3'].value
        if g3_value is not None and isinstance(g3_value, str) and g3_value.strip().startswith('='):
            print(f"PASS: Component 1 — G3 contains a formula: {repr(g3_value)} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — G3 should contain a formula, found: {repr(g3_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — could not read G3: {e}")

    # Component 2: G3 formula is COUNTIFS with correct ranges and criteria (0.6 points)
    # Verifies the formula references:
    #   - C2:C301 with criterion "Engineering"
    #   - D2:D301 with criterion ">90000"
    # This FAILS on initial (G3 is empty) and PASSES on golden
    try:
        g3_value = ws['G3'].value
        if g3_value is not None and isinstance(g3_value, str):
            # Normalize: remove spaces, uppercase for comparison
            formula_normalized = g3_value.upper().replace(' ', '')

            # Must be a COUNTIFS formula
            is_countifs = formula_normalized.startswith('=COUNTIFS(')

            # Must reference column C range for department check
            has_c_range = 'C2:C301' in formula_normalized

            # Must have "Engineering" as criterion (case-insensitive in formula)
            has_engineering = '"ENGINEERING"' in formula_normalized

            # Must reference column D range for salary check
            has_d_range = 'D2:D301' in formula_normalized

            # Must have ">90000" as salary criterion
            has_salary_criterion = '">90000"' in formula_normalized

            all_correct = is_countifs and has_c_range and has_engineering and has_d_range and has_salary_criterion

            if all_correct:
                print(f"PASS: Component 2 — G3 formula is correct COUNTIFS with proper ranges and criteria (0.6 pts)")
                print(f"  Formula: {repr(g3_value)}")
                total_score += 0.6
            else:
                missing = []
                if not is_countifs:
                    missing.append("not a COUNTIFS formula")
                if not has_c_range:
                    missing.append("missing C2:C301 range")
                if not has_engineering:
                    missing.append('missing "Engineering" criterion')
                if not has_d_range:
                    missing.append("missing D2:D301 range")
                if not has_salary_criterion:
                    missing.append('missing ">90000" criterion')
                print(f"FAIL: Component 2 — formula issues: {', '.join(missing)}")
                print(f"  Actual formula in G3: {repr(g3_value)}")
        else:
            print(f"FAIL: Component 2 — G3 is empty or not a formula: {repr(g3_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — could not verify COUNTIFS formula: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
