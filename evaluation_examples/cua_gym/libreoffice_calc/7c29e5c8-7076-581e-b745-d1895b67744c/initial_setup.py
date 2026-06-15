"""
Initial Setup: Create a spreadsheet with product revenue data for stacked area chart task
Task ID: calc_gg3_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Trends sheet ---
    ws = wb.active
    ws.title = 'Trends'

    # Headers
    headers = ['Month', 'Product A', 'Product B', 'Product C']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 6 months of realistic revenue data
    data = [
        ['January',   45230, 32150, 18760],
        ['February',  48900, 34800, 21340],
        ['March',     52100, 31200, 24500],
        ['April',     49750, 36900, 22800],
        ['May',       55300, 38400, 26100],
        ['June',      58200, 41500, 28900],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # NO chart in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
