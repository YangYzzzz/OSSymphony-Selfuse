"""
Reward Script: Singapore Restaurant Guide Task
Task ID: osworld_multi_apps_web_location_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): File exists with correct 3-sheet structure
  Component 2 (0.35): Sheet 1 "All Restaurants" has 15+ rows, all 9 required columns,
                       Distance_From_MBS_km values, and all 5 cuisine categories
  Component 3 (0.20): Sheet 2 "By Category" has per-category sections with restaurants
                       sorted by rating (descending) within each category
  Component 4 (0.15): Sheet 3 "Summary" has per-category aggregation rows
Total: 1.0
"""

import os

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_location_012'

# The golden file is saved as .xlsx (not .ods)
FILE_PATH_XLSX = f'{WORKDIR}/singapore_restaurant_guide.xlsx'
FILE_PATH_ODS  = f'{WORKDIR}/singapore_restaurant_guide.ods'

REQUIRED_COLUMNS = [
    'Name', 'Cuisine_Category', 'Address', 'Rating',
    'Price_Range', 'Signature_Dish', 'Latitude', 'Longitude', 'Distance_From_MBS_km'
]

REQUIRED_CATEGORIES = {'Chinese', 'Malay', 'Indian', 'Western', 'International'}

REQUIRED_SHEET1_NAMES = ['All Restaurants', 'all restaurants']
REQUIRED_SHEET2_NAMES = ['By Category', 'by category', 'ByCategory']
REQUIRED_SHEET3_NAMES = ['Summary', 'summary']


def normalize_headers(headers):
    """Normalize headers for loose comparison."""
    return [str(h).strip() if h is not None else '' for h in headers]


def find_sheet(wb, candidates):
    """Find a sheet by any of the candidate names (case-insensitive)."""
    lower_names = {n.lower(): n for n in wb.sheetnames}
    for candidate in candidates:
        if candidate.lower() in lower_names:
            return wb[lower_names[candidate.lower()]]
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Correct 3-sheet structure (0.30 points)
    # -----------------------------------------------------------------------
    try:
        sheet_count = len(wb.sheetnames)
        ws1 = find_sheet(wb, REQUIRED_SHEET1_NAMES)
        ws2 = find_sheet(wb, REQUIRED_SHEET2_NAMES)
        ws3 = find_sheet(wb, REQUIRED_SHEET3_NAMES)

        if ws1 is not None and ws2 is not None and ws3 is not None:
            print(f"PASS: Component 1 — 3 required sheets found: {wb.sheetnames} (0.30 pts)")
            total_score += 0.30
        else:
            missing = []
            if ws1 is None:
                missing.append("'All Restaurants'")
            if ws2 is None:
                missing.append("'By Category'")
            if ws3 is None:
                missing.append("'Summary'")
            print(f"FAIL: Component 1 — Missing sheets: {missing}. Found: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        ws1 = ws2 = ws3 = None

    # -----------------------------------------------------------------------
    # Component 2: Sheet 1 has 15+ restaurants with all required columns,
    #              Distance_From_MBS_km populated, and all 5 categories (0.35 points)
    # -----------------------------------------------------------------------
    if ws1 is not None:
        try:
            # Check headers
            headers = normalize_headers(
                [ws1.cell(row=1, column=c).value for c in range(1, ws1.max_column + 1)]
            )
            required_headers_found = all(col in headers for col in REQUIRED_COLUMNS)

            # Count data rows (non-empty, non-header)
            data_rows = []
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row, values_only=True):
                if any(v is not None for v in row):
                    data_rows.append(row)
            row_count = len(data_rows)

            # Check all 5 categories are present
            cuisine_col_idx = headers.index('Cuisine_Category') if 'Cuisine_Category' in headers else -1
            categories_found = set()
            if cuisine_col_idx >= 0:
                for row in data_rows:
                    if row[cuisine_col_idx]:
                        categories_found.add(str(row[cuisine_col_idx]).strip())
            all_categories_present = REQUIRED_CATEGORIES.issubset(categories_found)

            # Check Distance_From_MBS_km column is populated
            dist_col_idx = headers.index('Distance_From_MBS_km') if 'Distance_From_MBS_km' in headers else -1
            distances_populated = False
            if dist_col_idx >= 0:
                dist_values = [row[dist_col_idx] for row in data_rows if row[dist_col_idx] is not None]
                distances_populated = len(dist_values) >= 15

            # Also check Latitude and Longitude populated
            lat_col_idx = headers.index('Latitude') if 'Latitude' in headers else -1
            lng_col_idx = headers.index('Longitude') if 'Longitude' in headers else -1
            coords_populated = False
            if lat_col_idx >= 0 and lng_col_idx >= 0:
                coord_rows = [
                    row for row in data_rows
                    if row[lat_col_idx] is not None and row[lng_col_idx] is not None
                ]
                coords_populated = len(coord_rows) >= 15

            if (required_headers_found and row_count >= 15 and all_categories_present
                    and distances_populated and coords_populated):
                print(f"PASS: Component 2 — Sheet1 has {row_count} rows, all 9 columns present, "
                      f"categories: {sorted(categories_found)}, distances populated (0.35 pts)")
                total_score += 0.35
            else:
                issues = []
                if not required_headers_found:
                    missing_cols = [c for c in REQUIRED_COLUMNS if c not in headers]
                    issues.append(f"Missing columns: {missing_cols}")
                if row_count < 15:
                    issues.append(f"Only {row_count} data rows, need 15+")
                if not all_categories_present:
                    missing_cats = REQUIRED_CATEGORIES - categories_found
                    issues.append(f"Missing cuisine categories: {missing_cats}")
                if not distances_populated:
                    issues.append("Distance_From_MBS_km not fully populated")
                if not coords_populated:
                    issues.append("Latitude/Longitude not fully populated")
                print(f"FAIL: Component 2 — {'; '.join(issues)}")
        except Exception as e:
            print(f"ERROR: Component 2 — {e}")
    else:
        print("SKIP: Component 2 — Sheet 1 not found")

    # -----------------------------------------------------------------------
    # Component 3: Sheet 2 "By Category" has per-category sorted sections
    #              with restaurants sorted by rating descending (0.20 points)
    # -----------------------------------------------------------------------
    if ws2 is not None:
        try:
            # Parse By Category sheet: expect category headers with data sub-sections
            categories_seen = set()
            sort_violations = 0
            total_data_rows_cat = 0
            prev_rating = None
            in_data_section = False

            for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, values_only=True):
                row_name = row[0]
                # Skip empty rows
                if all(v is None for v in row):
                    prev_rating = None
                    in_data_section = False
                    continue
                # Skip sub-header rows (rows with 'Name' as first cell)
                if str(row_name).strip() == 'Name':
                    continue
                # Detect category header rows (no data in cols 2+)
                if all(v is None for v in row[1:]):
                    # This is a category header
                    prev_rating = None
                    in_data_section = False
                    # Extract category name from header like "Chinese Cuisine"
                    header_str = str(row_name).strip()
                    for cat in REQUIRED_CATEGORIES:
                        if cat.lower() in header_str.lower():
                            categories_seen.add(cat)
                            in_data_section = True
                            break
                    continue
                # Data row: check rating sorting
                if in_data_section and row[3] is not None:
                    try:
                        rating = float(row[3])
                        if prev_rating is not None and rating > prev_rating + 0.001:
                            sort_violations += 1
                        prev_rating = rating
                        total_data_rows_cat += 1
                    except (TypeError, ValueError):
                        pass

            all_categories_in_sheet2 = REQUIRED_CATEGORIES.issubset(categories_seen)
            sorting_correct = sort_violations == 0
            has_data = total_data_rows_cat >= 15

            if all_categories_in_sheet2 and sorting_correct and has_data:
                print(f"PASS: Component 3 — By Category sheet has all 5 categories, "
                      f"{total_data_rows_cat} data rows, sorted by rating desc (0.20 pts)")
                total_score += 0.20
            else:
                issues = []
                if not all_categories_in_sheet2:
                    missing_cats = REQUIRED_CATEGORIES - categories_seen
                    issues.append(f"Missing categories in Sheet2: {missing_cats}")
                if not sorting_correct:
                    issues.append(f"{sort_violations} sort violations (not sorted by rating desc)")
                if not has_data:
                    issues.append(f"Only {total_data_rows_cat} data rows in By Category")
                print(f"FAIL: Component 3 — {'; '.join(issues)}")
        except Exception as e:
            print(f"ERROR: Component 3 — {e}")
    else:
        print("SKIP: Component 3 — Sheet 2 not found")

    # -----------------------------------------------------------------------
    # Component 4: Sheet 3 "Summary" has per-category aggregation rows (0.15 points)
    # -----------------------------------------------------------------------
    if ws3 is not None:
        try:
            # Collect all non-empty rows in Summary sheet
            summary_rows = []
            for row in ws3.iter_rows(min_row=1, max_row=ws3.max_row, values_only=True):
                if any(v is not None for v in row):
                    summary_rows.append(row)

            # Must have at least 6 rows: 1 header + 5 category rows (+ optional OVERALL row)
            has_enough_rows = len(summary_rows) >= 6

            # Check that all 5 cuisine categories appear in column 1
            category_values = {str(r[0]).strip() for r in summary_rows if r[0] is not None}
            categories_in_summary = REQUIRED_CATEGORIES.issubset(category_values)

            # Check Avg_Rating and some numeric aggregation is present (col 2)
            numeric_values = []
            for row in summary_rows[1:]:  # skip header
                if row[1] is not None:
                    try:
                        numeric_values.append(float(row[1]))
                    except (TypeError, ValueError):
                        pass
            has_numeric_data = len(numeric_values) >= 5

            if has_enough_rows and categories_in_summary and has_numeric_data:
                print(f"PASS: Component 4 — Summary sheet has {len(summary_rows)} rows, "
                      f"all 5 categories present with numeric aggregation (0.15 pts)")
                total_score += 0.15
            else:
                issues = []
                if not has_enough_rows:
                    issues.append(f"Only {len(summary_rows)} rows in Summary (need 6+)")
                if not categories_in_summary:
                    missing_cats = REQUIRED_CATEGORIES - category_values
                    issues.append(f"Missing categories in Summary: {missing_cats}")
                if not has_numeric_data:
                    issues.append("Insufficient numeric aggregation values in Summary")
                print(f"FAIL: Component 4 — {'; '.join(issues)}")
        except Exception as e:
            print(f"ERROR: Component 4 — {e}")
    else:
        print("SKIP: Component 4 — Sheet 3 not found")

    # -----------------------------------------------------------------------
    # Final score
    # -----------------------------------------------------------------------
    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Guard: openpyxl must be available
if not OPENPYXL_AVAILABLE:
    print("CRITICAL: openpyxl is not installed on this VM — cannot verify spreadsheet.")
    print("REWARD: 0.0")
# Determine which file to verify (support both .xlsx and .ods)
elif os.path.exists(FILE_PATH_XLSX):
    verify_task(FILE_PATH_XLSX)
elif os.path.exists(FILE_PATH_ODS):
    verify_task(FILE_PATH_ODS)
else:
    print(f"File not found: {FILE_PATH_XLSX} or {FILE_PATH_ODS}")
    print("REWARD: 0.0")
