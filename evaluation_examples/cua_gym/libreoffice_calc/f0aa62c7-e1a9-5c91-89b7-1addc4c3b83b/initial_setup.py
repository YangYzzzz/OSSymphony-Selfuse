"""
Initial Setup: Apply data validation to cells B2:B30 with dropdown list
Task ID: calc_nrv_086
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Main transaction sheet ---
    ws1 = wb.active
    ws1.title = "Sheet1"

    # Headers
    ws1.cell(row=1, column=1, value="Transaction")
    ws1.cell(row=1, column=2, value="Category")

    # Style headers
    header_font = Font(bold=True, size=11)
    for col in [1, 2]:
        ws1.cell(row=1, column=col).font = header_font

    # Realistic transaction data in column A (B2:B30 left empty - no validation)
    transactions = [
        "Office supplies purchase - Mar 2025",
        "Client dinner at Riverside Grill",
        "Monthly software subscription",
        "Flight to Chicago conference",
        "Hotel stay - 3 nights Portland",
        "Uber rides - week of Mar 10",
        "Team building event catering",
        "New laptop for design team",
        "Quarterly insurance premium",
        "Marketing brochure printing",
        "Server hosting renewal",
        "Training workshop registration",
        "Employee wellness program",
        "Legal consultation fee",
        "Office furniture replacement",
        "Cloud storage upgrade",
        "Annual trade show booth",
        "Courier and shipping charges",
        "Building maintenance repair",
        "Professional photography session",
        "Accounting software license",
        "Safety equipment purchase",
        "Holiday party venue rental",
        "Consulting engagement - Phase 2",
        "IT security audit",
        "Recruitment agency fee",
        "Company vehicle fuel",
        "Charitable donation - local shelter",
        "Equipment calibration service",
    ]
    for r, txn in enumerate(transactions, 2):
        ws1.cell(row=r, column=1, value=txn)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 42
    ws1.column_dimensions['B'].width = 22

    # --- Sheet 2: Categories reference sheet ---
    ws2 = wb.create_sheet("Categories")
    ws2.cell(row=1, column=1, value="Category Name")
    ws2.cell(row=1, column=1).font = Font(bold=True, size=11)

    categories = [
        "Office Supplies",
        "Meals & Entertainment",
        "Software & Subscriptions",
        "Travel - Airfare",
        "Travel - Lodging",
        "Transportation",
        "Events & Catering",
        "Equipment & Hardware",
        "Insurance",
        "Marketing & Advertising",
        "IT & Infrastructure",
        "Training & Development",
        "Health & Wellness",
        "Professional Services",
    ]
    for r, cat in enumerate(categories, 2):
        ws2.cell(row=r, column=1, value=cat)

    ws2.column_dimensions['A'].width = 28

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
