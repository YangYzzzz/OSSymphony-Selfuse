"""
Initial Setup: Customer health score matrix with raw data
Task ID: calc_sales_080
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Health ---
    ws = wb.active
    ws.title = 'Health'

    # Headers
    headers = [
        'Customer', 'Usage %', 'Support Tickets (90d)', 'NPS',
        'Days to Renewal', 'Usage Score', 'Support Score',
        'NPS Score', 'Renewal Urgency', 'Health Score', 'Status'
    ]
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows (columns A-E only; F-K left empty for the agent to fill)
    data = [
        ['Cust A', 0.85, 2, 9, 45],
        ['Cust B', 0.40, 8, 5, 180],
        ['Cust C', 0.72, 1, 8, 30],
        ['Cust D', 0.95, 0, 10, 300],
        ['Cust E', 0.25, 12, 3, 60],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 2:  # Usage % column - format as percentage
                cell.number_format = '0%'
            elif c == 5:  # Days to Renewal
                cell.number_format = '0'

    # Set column widths for readability
    col_widths = {
        'A': 14, 'B': 12, 'C': 22, 'D': 8, 'E': 18,
        'F': 14, 'G': 15, 'H': 12, 'I': 17, 'J': 14, 'K': 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row height for header
    ws.row_dimensions[1].height = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
