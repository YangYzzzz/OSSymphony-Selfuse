"""
Reward Script: Customer health score matrix with composite health indicator
Task ID: calc_sales_080
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Usage Score formulas in F2:F6
  Component 2 (0.25): Support Score (G) and NPS Score (H) formulas in rows 2-6
  Component 3 (0.25): Renewal Urgency (I) and Health Score (J) formulas in rows 2-6
  Component 4 (0.25): Status classification (K) formulas in rows 2-6
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_080'

# Ground truth expected values (computed from task context)
# F: Usage Score  G: Support Score  H: NPS Score  I: Renewal Urgency  J: Health Score  K: Status
EXPECTED = {
    2: {'F': 5, 'G': 4, 'H': 5, 'I': 4, 'J': 90, 'K': 'Healthy'},
    3: {'F': 3, 'G': 2, 'H': 3, 'I': 2, 'J': 50, 'K': 'At Risk'},
    4: {'F': 4, 'G': 5, 'H': 4, 'I': 5, 'J': 88, 'K': 'Healthy'},
    5: {'F': 5, 'G': 5, 'H': 5, 'I': 1, 'J': 84, 'K': 'Healthy'},
    6: {'F': 2, 'G': 1, 'H': 2, 'I': 4, 'J': 42, 'K': 'At Risk'},
}


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def is_valid_if_formula(val):
    """Check if a cell value is an IF-based formula (not None/empty)."""
    if not isinstance(val, str):
        return False
    return val.strip().startswith('=') and 'IF(' in val.upper()


def compute_usage_score(usage_pct):
    """Compute expected usage score from raw usage percentage."""
    if usage_pct >= 0.8: return 5
    if usage_pct >= 0.6: return 4
    if usage_pct >= 0.4: return 3
    if usage_pct >= 0.2: return 2
    return 1


def compute_support_score(tickets):
    """Compute expected support score from ticket count."""
    if tickets <= 1: return 5
    if tickets <= 3: return 4
    if tickets <= 5: return 3
    if tickets <= 8: return 2
    return 1


def compute_nps_score(nps):
    """Compute expected NPS score."""
    if nps >= 9: return 5
    if nps >= 7: return 4
    if nps >= 5: return 3
    if nps >= 3: return 2
    return 1


def compute_renewal_urgency(days):
    """Compute expected renewal urgency score."""
    if days <= 30: return 5
    if days <= 60: return 4
    if days <= 90: return 3
    if days <= 180: return 2
    return 1


def compute_health_score(f_val, g_val, h_val, i_val):
    """Compute expected health score from component scores."""
    return (f_val * 0.3 + g_val * 0.25 + h_val * 0.25 + i_val * 0.2) * 20


def compute_status(health_score):
    """Compute expected status from health score."""
    if health_score >= 80: return 'Healthy'
    if health_score >= 60: return 'Monitor'
    return 'At Risk'


def check_formula_or_value(ws, ws_data, col_letter, row, expected_val, tolerance=0.5):
    """
    Check if a cell has either:
    1. A correct formula (IF-based for F-I, composite for J, IF-based for K), OR
    2. The correct computed value (if formulas were evaluated by LibreOffice)
    Returns True if either condition is met.
    """
    cell_ref = f'{col_letter}{row}'
    formula_val = ws[cell_ref].value
    cached_val = ws_data[cell_ref].value

    # Check 1: Is there a formula present?
    if isinstance(formula_val, str) and formula_val.startswith('='):
        # For columns F-I and K: should be IF-based
        if col_letter in ('F', 'G', 'H', 'I', 'K'):
            if 'IF(' in formula_val.upper():
                # Verify the formula references the correct source column
                row_ref = str(row)
                if col_letter == 'F' and f'B{row_ref}' in formula_val.upper():
                    return True
                elif col_letter == 'G' and f'C{row_ref}' in formula_val.upper():
                    return True
                elif col_letter == 'H' and f'D{row_ref}' in formula_val.upper():
                    return True
                elif col_letter == 'I' and f'E{row_ref}' in formula_val.upper():
                    return True
                elif col_letter == 'K' and f'J{row_ref}' in formula_val.upper():
                    return True
        # For column J: should be a weighted composite formula
        elif col_letter == 'J':
            upper = formula_val.upper()
            # Check it references F, G, H, I columns for this row
            if (f'F{row}' in upper and f'G{row}' in upper and
                f'H{row}' in upper and f'I{row}' in upper):
                return True

    # Check 2: Is the cached/computed value correct?
    if cached_val is not None:
        if col_letter == 'K':
            if isinstance(cached_val, str) and cached_val.strip() == expected_val:
                return True
        else:
            try:
                if abs(float(cached_val) - float(expected_val)) <= tolerance:
                    return True
            except (ValueError, TypeError):
                pass

    # Check 3: Maybe the formula was evaluated and stored as a direct value
    if not isinstance(formula_val, str):
        if formula_val is not None:
            if col_letter == 'K':
                if isinstance(formula_val, str) and formula_val.strip() == expected_val:
                    return True
            else:
                try:
                    if abs(float(formula_val) - float(expected_val)) <= tolerance:
                        return True
                except (ValueError, TypeError):
                    pass

    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"WARNING: Cannot load data_only workbook: {e}")
        wb_data = wb

    # Check sheet exists
    if 'Health' not in wb.sheetnames:
        print("CRITICAL: 'Health' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Health']
    ws_data = wb_data['Health']

    # Component 1: Usage Score formulas F2:F6 (0.25 points)
    # These cells should be None/empty in initial, populated in golden
    try:
        usage_pass = 0
        for row in range(2, 7):
            expected = EXPECTED[row]['F']
            if check_formula_or_value(ws, ws_data, 'F', row, expected):
                usage_pass += 1
                print(f"  PASS: F{row} has correct usage score formula/value (expected {expected})")
            else:
                actual = ws[f'F{row}'].value
                print(f"  FAIL: F{row} expected usage score {expected}, found: {repr(actual)}")

        if usage_pass == 5:
            print(f"PASS: Component 1 -- All 5 Usage Score formulas correct (0.25 pts)")
            total_score += 0.25
        elif usage_pass >= 3:
            partial = round(0.25 * usage_pass / 5, 2)
            print(f"PARTIAL: Component 1 -- {usage_pass}/5 Usage Score formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {usage_pass}/5 Usage Score formulas correct")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Support Score (G) and NPS Score (H) formulas (0.25 points)
    try:
        gh_pass = 0
        total_gh = 10  # 5 rows x 2 columns
        for row in range(2, 7):
            for col in ('G', 'H'):
                expected = EXPECTED[row][col]
                if check_formula_or_value(ws, ws_data, col, row, expected):
                    gh_pass += 1
                    print(f"  PASS: {col}{row} has correct formula/value (expected {expected})")
                else:
                    actual = ws[f'{col}{row}'].value
                    print(f"  FAIL: {col}{row} expected {expected}, found: {repr(actual)}")

        if gh_pass == total_gh:
            print(f"PASS: Component 2 -- All {total_gh} Support/NPS Score formulas correct (0.25 pts)")
            total_score += 0.25
        elif gh_pass >= 6:
            partial = round(0.25 * gh_pass / total_gh, 2)
            print(f"PARTIAL: Component 2 -- {gh_pass}/{total_gh} formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {gh_pass}/{total_gh} Support/NPS formulas correct")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Renewal Urgency (I) and Health Score (J) formulas (0.25 points)
    try:
        ij_pass = 0
        total_ij = 10  # 5 rows x 2 columns
        for row in range(2, 7):
            for col in ('I', 'J'):
                expected = EXPECTED[row][col]
                if check_formula_or_value(ws, ws_data, col, row, expected):
                    ij_pass += 1
                    print(f"  PASS: {col}{row} has correct formula/value (expected {expected})")
                else:
                    actual = ws[f'{col}{row}'].value
                    print(f"  FAIL: {col}{row} expected {expected}, found: {repr(actual)}")

        if ij_pass == total_ij:
            print(f"PASS: Component 3 -- All {total_ij} Renewal/Health formulas correct (0.25 pts)")
            total_score += 0.25
        elif ij_pass >= 6:
            partial = round(0.25 * ij_pass / total_ij, 2)
            print(f"PARTIAL: Component 3 -- {ij_pass}/{total_ij} formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Only {ij_pass}/{total_ij} Renewal/Health formulas correct")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Status classification (K) formulas (0.25 points)
    try:
        k_pass = 0
        for row in range(2, 7):
            expected = EXPECTED[row]['K']
            if check_formula_or_value(ws, ws_data, 'K', row, expected):
                k_pass += 1
                print(f"  PASS: K{row} has correct status formula/value (expected '{expected}')")
            else:
                actual = ws[f'K{row}'].value
                print(f"  FAIL: K{row} expected status '{expected}', found: {repr(actual)}")

        if k_pass == 5:
            print(f"PASS: Component 4 -- All 5 Status formulas correct (0.25 pts)")
            total_score += 0.25
        elif k_pass >= 3:
            partial = round(0.25 * k_pass / 5, 2)
            print(f"PARTIAL: Component 4 -- {k_pass}/5 Status formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- Only {k_pass}/5 Status formulas correct")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state before verification
def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


# Main execution
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state("libreoffice_calc")
    verify_task(file_path)
