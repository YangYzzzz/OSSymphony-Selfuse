"""
Initial Setup: Churn risk scoring model — customer activity spreadsheet
Task ID: calc_sales_customer_churn_023
Domain: libreoffice_calc

Creates a CustomerActivity sheet with 300 customers.
Columns D (Days Since Purchase), E (Churn Score), F (Risk Status) are EMPTY
because the task asks the agent to fill them in.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random
import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_customer_churn_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

# Realistic company names
COMPANIES = [
    "Apex Solutions", "BlueSky Technologies", "Cascade Analytics", "Delta Dynamics",
    "Ember Systems", "Fusion Labs", "Granite Partners", "Harbor Networks",
    "Irongate Consulting", "Jade Innovations", "Keystone Digital", "Luminary Group",
    "Meridian Corp", "Nexus Ventures", "Orbit Strategies", "Pinnacle Services",
    "Quasar Media", "Redwood Enterprises", "Summit Analytics", "Titan Industries",
    "Unified Tech", "Vertex Solutions", "Westfield Capital", "Xcel Dynamics",
    "Yellowstone Group", "Zenith Partners", "Acme Corp", "Benchmark Systems",
    "Cobalt Technologies", "Driftwood Media", "Eclipse Ventures", "Fulcrum Group",
    "Greenlight Partners", "Highpoint Digital", "Integral Analytics", "Junction Labs",
    "Kestrel Networks", "Lighthouse Consulting", "Mosaic Solutions", "Nordic Capital",
    "Onyx Enterprises", "Parallax Group", "Quantum Strategies", "Ridgeline Services",
    "Strata Systems", "Terranova Analytics", "Uptown Ventures", "Vanguard Media",
    "Windmill Technologies", "Xenolith Partners",
]

# Distribution of last purchase dates to ensure variety in churn scores
# Approx 25% each category
def get_last_purchase_date(idx):
    today = datetime.date(2026, 3, 4)
    # Distribute across 5 bands for variety
    bucket = idx % 5
    if bucket == 0:
        # Under 30 days (score 1)
        days_ago = random.randint(1, 29)
    elif bucket == 1:
        # 31-60 days (score 2)
        days_ago = random.randint(31, 60)
    elif bucket == 2:
        # 61-90 days (score 3)
        days_ago = random.randint(61, 90)
    elif bucket == 3:
        # 91-180 days (score 4) — "At Risk"
        days_ago = random.randint(91, 180)
    else:
        # Over 180 days (score 5) — "At Risk"
        days_ago = random.randint(181, 365)
    return today - datetime.timedelta(days=days_ago)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CustomerActivity'

    # --- Headers ---
    headers = ['Customer ID', 'Company', 'Last Purchase Date', 'Days Since Purchase',
               'Churn Score', 'Risk Status']
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    # --- Data rows (300 customers) ---
    for i in range(1, 301):
        row = i + 1  # data starts at row 2
        company = COMPANIES[(i - 1) % len(COMPANIES)]
        last_purchase = get_last_purchase_date(i - 1)

        # Customer ID
        ws.cell(row=row, column=1, value=f'CUST-{1000 + i:04d}')
        # Company
        ws.cell(row=row, column=2, value=company)
        # Last Purchase Date — store as date value
        date_cell = ws.cell(row=row, column=3, value=last_purchase)
        date_cell.number_format = 'YYYY-MM-DD'

        # Columns D, E, F are intentionally left EMPTY
        # (agent will fill Days Since Purchase, Churn Score, Risk Status)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: CustomerActivity')
    print(f'  Rows: 301 (1 header + 300 data)')
    print(f'  Columns D, E, F: empty (to be filled by agent)')


create_initial()
