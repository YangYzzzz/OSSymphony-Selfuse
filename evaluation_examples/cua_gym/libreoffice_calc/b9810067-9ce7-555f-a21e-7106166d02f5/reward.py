"""
Reward Script: Build a churn risk scoring model
Task ID: calc_sales_customer_churn_023
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Days Since Purchase formulas in D2:D301 (=TODAY()-Cx)       — 0.25 pts
  Component 2: Churn Score IFS formulas in E2:E301 with correct thresholds — 0.30 pts
  Component 3: Risk Status IF formulas in F2:F301 (score >= 4 => At Risk)  — 0.20 pts
  Component 4: Conditional formatting on F2:F301 (red/green fills)         — 0.10 pts
  Component 5: Data sorted with At Risk customers first                    — 0.15 pts
  Total: 1.0
"""

import os
import openpyxl
from datetime import datetime

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_sales_customer_churn_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheet exists
    if 'CustomerActivity' not in wb.sheetnames:
        print("CRITICAL: Sheet 'CustomerActivity' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CustomerActivity']

    # Component 1: Days Since Purchase formulas in D2:D301 (0.25 points)
    # Each cell should contain a formula like =TODAY()-C<row>
    # This FAILS on initial (D column is empty) and PASSES on golden
    try:
        d_formula_count = 0
        d_correct_count = 0
        total_d_rows = 300  # rows 2-301

        for row in range(2, 302):
            d_val = ws.cell(row=row, column=4).value
            if d_val is not None:
                d_formula_count += 1
                # Check it's a formula referencing TODAY() and the C column for that row
                if isinstance(d_val, str) and d_val.startswith('='):
                    d_upper = d_val.upper()
                    # Should reference TODAY() and C<row>
                    if 'TODAY()' in d_upper and f'C{row}' in d_val:
                        d_correct_count += 1

        if d_correct_count == total_d_rows:
            print(f"PASS: Component 1 — All {total_d_rows} D column cells have correct =TODAY()-C<row> formulas (0.25 pts)")
            total_score += 0.25
        elif d_formula_count == total_d_rows:
            # Formulas exist but pattern may differ slightly
            print(f"PASS (partial): Component 1 — All {total_d_rows} D column cells have formulas, but {d_correct_count}/{total_d_rows} match exactly. Granting partial credit.")
            # Check at least majority have today-date formulas
            if d_correct_count >= total_d_rows * 0.9:
                total_score += 0.25
            elif d_correct_count >= total_d_rows * 0.5:
                total_score += 0.15
            else:
                total_score += 0.0
        else:
            print(f"FAIL: Component 1 — Only {d_formula_count}/{total_d_rows} D column cells are filled. Expected =TODAY()-C<row> formulas.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Churn Score IFS formulas in E2:E301 (0.30 points)
    # Each cell should have =IFS(D<row><30,1,D<row><=60,2,D<row><=90,3,D<row><=180,4,D<row>>180,5)
    # This FAILS on initial (E column is empty) and PASSES on golden
    try:
        e_formula_count = 0
        e_correct_count = 0
        e_has_ifs_count = 0
        total_e_rows = 300

        for row in range(2, 302):
            e_val = ws.cell(row=row, column=5).value
            if e_val is not None:
                e_formula_count += 1
                if isinstance(e_val, str) and e_val.startswith('='):
                    e_upper = e_val.upper()
                    # Check for IFS formula with the required thresholds (30, 60, 90, 180)
                    if 'IFS(' in e_upper or 'IFS(' in e_val.upper():
                        e_has_ifs_count += 1
                        # Must check all 5 threshold conditions
                        if '<30' in e_upper and '<=60' in e_upper and '<=90' in e_upper and '<=180' in e_upper and '>180' in e_upper:
                            e_correct_count += 1

        if e_correct_count == total_e_rows:
            print(f"PASS: Component 2 — All {total_e_rows} E column cells have correct IFS formulas with thresholds 30/60/90/180 (0.30 pts)")
            total_score += 0.30
        elif e_has_ifs_count == total_e_rows:
            print(f"PASS (partial): Component 2 — All cells have IFS formulas but {e_correct_count}/{total_e_rows} have all correct thresholds.")
            if e_correct_count >= total_e_rows * 0.9:
                total_score += 0.25
            elif e_correct_count >= total_e_rows * 0.5:
                total_score += 0.15
            else:
                total_score += 0.10
        elif e_formula_count == total_e_rows:
            print(f"FAIL (partial): Component 2 — {e_formula_count} E cells filled with formulas but only {e_has_ifs_count} use IFS. Found {e_correct_count} correct ones.")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — Only {e_formula_count}/{total_e_rows} E column cells are filled. Expected IFS formulas with score thresholds.")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Risk Status IF formulas in F2:F301 (0.20 points)
    # Each cell should have =IF(E<row>>=4,"At Risk","Healthy")
    # This FAILS on initial (F column is empty) and PASSES on golden
    try:
        f_formula_count = 0
        f_correct_count = 0
        total_f_rows = 300

        for row in range(2, 302):
            f_val = ws.cell(row=row, column=6).value
            if f_val is not None:
                f_formula_count += 1
                if isinstance(f_val, str) and f_val.startswith('='):
                    f_upper = f_val.upper()
                    # Check for IF formula with E>=4, "At Risk", "Healthy"
                    if 'IF(' in f_upper:
                        # E<row>>=4 and mentions "AT RISK" and "HEALTHY"
                        if f'E{row}>=4' in f_val or f'E{row}>=4' in f_upper:
                            if 'AT RISK' in f_upper and 'HEALTHY' in f_upper:
                                f_correct_count += 1

        if f_correct_count == total_f_rows:
            print(f"PASS: Component 3 — All {total_f_rows} F column cells have correct =IF(E>=4,\"At Risk\",\"Healthy\") formulas (0.20 pts)")
            total_score += 0.20
        elif f_formula_count == total_f_rows:
            print(f"PASS (partial): Component 3 — All cells have formulas but only {f_correct_count}/{total_f_rows} exactly match. Awarding partial.")
            if f_correct_count >= total_f_rows * 0.9:
                total_score += 0.18
            elif f_correct_count >= total_f_rows * 0.5:
                total_score += 0.12
            else:
                total_score += 0.05
        else:
            print(f"FAIL: Component 3 — Only {f_formula_count}/{total_f_rows} F column cells are filled. Expected =IF(E>=4,\"At Risk\",\"Healthy\") formulas.")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on F2:F301 (0.10 points)
    # Red fill for 'At Risk', green fill for 'Healthy'
    # This FAILS on initial (no conditional formatting) and PASSES on golden
    try:
        cf_present = False
        has_red_for_at_risk = False
        has_green_for_healthy = False

        for cf_range, cf_rules in ws.conditional_formatting._cf_rules.items():
            cf_range_str = str(cf_range)
            # Check if the conditional formatting applies to F column rows
            if 'F' in cf_range_str:
                for rule in cf_rules:
                    cf_present = True
                    formula = getattr(rule, 'formula', None)
                    if formula:
                        formula_upper = ' '.join(formula).upper()
                        dxf = getattr(rule, 'dxf', None)
                        fill = getattr(dxf, 'fill', None) if dxf else None

                        if 'AT RISK' in formula_upper and fill:
                            try:
                                fg_rgb = fill.fgColor.rgb
                                # Red is FFFF0000
                                if fg_rgb and 'FF0000' in fg_rgb.upper():
                                    has_red_for_at_risk = True
                            except Exception:
                                pass

                        if 'HEALTHY' in formula_upper and fill:
                            try:
                                fg_rgb = fill.fgColor.rgb
                                # Green is FF00FF00
                                if fg_rgb and '00FF00' in fg_rgb.upper():
                                    has_green_for_healthy = True
                            except Exception:
                                pass

        if cf_present and has_red_for_at_risk and has_green_for_healthy:
            print(f"PASS: Component 4 — Conditional formatting on F column: red for 'At Risk', green for 'Healthy' (0.10 pts)")
            total_score += 0.10
        elif cf_present and (has_red_for_at_risk or has_green_for_healthy):
            print(f"PASS (partial): Component 4 — Conditional formatting present but missing one rule. red={has_red_for_at_risk}, green={has_green_for_healthy}")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on F column. Expected red='At Risk', green='Healthy'.")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Data sorted with At Risk customers first (0.15 points)
    # At Risk customers (score >= 4, days >= 91) should appear before Healthy customers
    # This FAILS on initial (data unsorted, D/E/F empty) and PASSES on golden
    try:
        today = datetime.now()

        # Compute expected score for each row based on C column (Last Purchase Date)
        # Find first Healthy customer row and check if any At Risk rows come after it
        first_healthy_row = None
        at_risk_after_healthy = False
        valid_rows = 0

        for row in range(2, 302):
            c_val = ws.cell(row=row, column=3).value
            if c_val and isinstance(c_val, datetime):
                valid_rows += 1
                days = (today - c_val).days
                # Score >= 4 means at risk (91-180 days = score 4, >180 days = score 5)
                is_at_risk = days >= 91

                if not is_at_risk:
                    if first_healthy_row is None:
                        first_healthy_row = row
                else:
                    # At Risk customer
                    if first_healthy_row is not None:
                        at_risk_after_healthy = True
                        break  # Found an at-risk after healthy — unsorted

        if valid_rows == 0:
            print("FAIL: Component 5 — No valid date data found in C column")
        elif first_healthy_row is None:
            print(f"INFO: Component 5 — All {valid_rows} rows appear to be At Risk (score >= 4)")
            total_score += 0.15
        elif at_risk_after_healthy:
            print(f"FAIL: Component 5 — Data not sorted: found At Risk customer after Healthy customer (first Healthy at row {first_healthy_row})")
        else:
            print(f"PASS: Component 5 — Data sorted with At Risk customers first (Healthy starts at row {first_healthy_row}) (0.15 pts)")
            total_score += 0.15
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
