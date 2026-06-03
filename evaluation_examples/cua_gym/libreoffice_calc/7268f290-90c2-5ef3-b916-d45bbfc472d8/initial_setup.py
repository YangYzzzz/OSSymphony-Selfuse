"""
Initial Setup: INDIRECT dynamic cell reference from monthly sales sheets
Task ID: calc_fma_indirect_dynamic_sheet_076
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_indirect_dynamic_sheet_076'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # ---- Monthly sheets: Jan, Feb, Mar, Apr, May, Jun ----
    # Each has product info in columns A-C and sales values in column B rows 5-15

    monthly_data = {
        'Jan': [
            # (row, col_a_product, col_b_sales, col_c_units)
            (1, 'Product', 'Sales ($)', 'Units'),
            (2, 'Alpha Widget', 11200, 45),
            (3, 'Beta Gadget', 8750, 30),
            (4, 'Gamma Device', 15400, 62),
            (5, 'Delta Tool', 22300, 89),
            (6, 'Epsilon Part', 9800, 40),
            (7, 'Zeta Component', 31500, 126),
            (8, 'Eta Module', 18900, 76),
            (9, 'Theta Unit', 27450, 110),
            (10, 'Iota System', 14200, 57),
            (11, 'Kappa Bundle', 39800, 159),
            (12, 'Lambda Pack', 21100, 84),
            (13, 'Mu Assembly', 17600, 70),
            (14, 'Nu Kit', 12500, 50),
            (15, 'Xi Set', 28700, 115),
        ],
        'Feb': [
            (1, 'Product', 'Sales ($)', 'Units'),
            (2, 'Alpha Widget', 10900, 43),
            (3, 'Beta Gadget', 9200, 37),
            (4, 'Gamma Device', 16100, 65),
            (5, 'Delta Tool', 19800, 79),
            (6, 'Epsilon Part', 11300, 45),
            (7, 'Zeta Component', 29700, 119),
            (8, 'Eta Module', 24600, 98),
            (9, 'Theta Unit', 30150, 121),
            (10, 'Iota System', 13700, 55),
            (11, 'Kappa Bundle', 42300, 169),
            (12, 'Lambda Pack', 18900, 76),
            (13, 'Mu Assembly', 21500, 86),
            (14, 'Nu Kit', 14800, 59),
            (15, 'Xi Set', 26300, 105),
        ],
        'Mar': [
            (1, 'Product', 'Sales ($)', 'Units'),
            (2, 'Alpha Widget', 13400, 54),
            (3, 'Beta Gadget', 10200, 41),
            (4, 'Gamma Device', 17300, 69),
            (5, 'Delta Tool', 24700, 99),
            (6, 'Epsilon Part', 8900, 36),
            (7, 'Zeta Component', 33800, 135),
            (8, 'Eta Module', 20100, 80),
            (9, 'Theta Unit', 25800, 103),
            (10, 'Iota System', 15600, 62),
            (11, 'Kappa Bundle', 37500, 150),
            (12, 'Lambda Pack', 22400, 90),
            (13, 'Mu Assembly', 19200, 77),
            (14, 'Nu Kit', 16700, 67),
            (15, 'Xi Set', 30100, 120),
        ],
        'Apr': [
            (1, 'Product', 'Sales ($)', 'Units'),
            (2, 'Alpha Widget', 12100, 48),
            (3, 'Beta Gadget', 9700, 39),
            (4, 'Gamma Device', 14900, 60),
            (5, 'Delta Tool', 21500, 86),
            (6, 'Epsilon Part', 10600, 42),
            (7, 'Zeta Component', 28400, 114),
            (8, 'Eta Module', 17800, 71),
            (9, 'Theta Unit', 32100, 128),
            (10, 'Iota System', 16300, 65),
            (11, 'Kappa Bundle', 44200, 177),
            (12, 'Lambda Pack', 19700, 79),
            (13, 'Mu Assembly', 23100, 92),
            (14, 'Nu Kit', 11900, 48),
            (15, 'Xi Set', 27600, 110),
        ],
        'May': [
            (1, 'Product', 'Sales ($)', 'Units'),
            (2, 'Alpha Widget', 14700, 59),
            (3, 'Beta Gadget', 11500, 46),
            (4, 'Gamma Device', 18600, 74),
            (5, 'Delta Tool', 26800, 107),
            (6, 'Epsilon Part', 12100, 48),
            (7, 'Zeta Component', 35200, 141),
            (8, 'Eta Module', 22300, 89),
            (9, 'Theta Unit', 29400, 118),
            (10, 'Iota System', 18900, 76),
            (11, 'Kappa Bundle', 41700, 167),
            (12, 'Lambda Pack', 24500, 98),
            (13, 'Mu Assembly', 20600, 82),
            (14, 'Nu Kit', 17300, 69),
            (15, 'Xi Set', 31800, 127),
        ],
        'Jun': [
            (1, 'Product', 'Sales ($)', 'Units'),
            (2, 'Alpha Widget', 15200, 61),
            (3, 'Beta Gadget', 12400, 50),
            (4, 'Gamma Device', 19800, 79),
            (5, 'Delta Tool', 28100, 112),
            (6, 'Epsilon Part', 13500, 54),
            (7, 'Zeta Component', 36900, 148),
            (8, 'Eta Module', 23800, 95),
            (9, 'Theta Unit', 31600, 126),
            (10, 'Iota System', 20300, 81),
            (11, 'Kappa Bundle', 46500, 186),
            (12, 'Lambda Pack', 26200, 105),
            (13, 'Mu Assembly', 22400, 90),
            (14, 'Nu Kit', 18700, 75),
            (15, 'Xi Set', 33400, 134),
        ],
    }

    # Use the first default sheet as Jan
    ws_jan = wb.active
    ws_jan.title = 'Jan'

    month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun']

    for i, month in enumerate(month_order):
        if i == 0:
            ws = ws_jan
        else:
            ws = wb.create_sheet(month)

        # Bold header row
        ws.cell(row=1, column=1, value='Product').font = Font(bold=True)
        ws.cell(row=1, column=2, value='Sales ($)').font = Font(bold=True)
        ws.cell(row=1, column=3, value='Units').font = Font(bold=True)

        for (row, col_a, col_b, col_c) in monthly_data[month]:
            if row == 1:
                continue  # headers already written
            ws.cell(row=row, column=1, value=col_a)
            ws.cell(row=row, column=2, value=col_b)
            ws.cell(row=row, column=3, value=col_c)

        ws.column_dimensions['A'].width = 18
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10

    # ---- Summary sheet ----
    ws_sum = wb.create_sheet('Summary')

    # Headers
    ws_sum.cell(row=1, column=1, value='Month').font = Font(bold=True)
    ws_sum.cell(row=1, column=2, value='Row').font = Font(bold=True)
    ws_sum.cell(row=1, column=3, value='Retrieved Value').font = Font(bold=True)

    # Data rows 2-9: month names and row numbers from task context
    summary_data = [
        ('Jan', 5),
        ('Feb', 7),
        ('Mar', 9),
        ('Jan', 11),
        ('Apr', 5),
        ('Feb', 8),
        ('Mar', 6),
        ('May', 10),
    ]

    for r, (month_name, row_num) in enumerate(summary_data, 2):
        ws_sum.cell(row=r, column=1, value=month_name)
        ws_sum.cell(row=r, column=2, value=row_num)
        # Column C (Retrieved Value) intentionally left empty

    ws_sum.column_dimensions['A'].width = 10
    ws_sum.column_dimensions['B'].width = 8
    ws_sum.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
