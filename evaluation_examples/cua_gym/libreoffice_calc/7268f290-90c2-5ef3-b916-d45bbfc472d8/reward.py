"""
Reward Script: Use INDIRECT to dynamically construct cell references and retrieve values
Task ID: calc_fma_indirect_dynamic_sheet_076
Domain: libreoffice_calc
Scoring:
  Component 1: All 8 cells C2:C9 contain INDIRECT formulas (0.6 points)
  Component 2: Formulas match the exact expected pattern =INDIRECT(An&".B"&Bn) (0.25 points)
  Component 3: No other Summary cells were modified (headers and A/B columns intact) (0.15 points)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_indirect_dynamic_sheet_076'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires adding =INDIRECT(An&".B"&Bn) formulas to cells C2:C9
    in the 'Summary' sheet to dynamically retrieve sales values from monthly sheets.
    Initial state: C2:C9 are all empty (None).
    Golden state:  C2:C9 each contain a formula like =INDIRECT(A2&".B"&B2).
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify Summary sheet exists (precondition gate)
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Expected INDIRECT formulas for C2:C9
    # Row n: =INDIRECT(An&".B"&Bn)
    expected_formulas = {
        2: '=INDIRECT(A2&".B"&B2)',
        3: '=INDIRECT(A3&".B"&B3)',
        4: '=INDIRECT(A4&".B"&B4)',
        5: '=INDIRECT(A5&".B"&B5)',
        6: '=INDIRECT(A6&".B"&B6)',
        7: '=INDIRECT(A7&".B"&B7)',
        8: '=INDIRECT(A8&".B"&B8)',
        9: '=INDIRECT(A9&".B"&B9)',
    }

    # Component 1: All 8 cells C2:C9 contain INDIRECT formulas (0.6 points)
    # This FAILS on initial (all None) and should PASS on golden.
    try:
        cells_with_indirect = 0
        cells_missing = []
        for row in range(2, 10):
            val = ws.cell(row=row, column=3).value
            if val is not None and isinstance(val, str) and 'INDIRECT' in val.upper():
                cells_with_indirect += 1
            else:
                cells_missing.append(f"C{row}={repr(val)}")

        if cells_with_indirect == 8:
            print(f"PASS: Component 1 — All 8 cells C2:C9 contain INDIRECT formulas (0.6 pts)")
            total_score += 0.6
        elif cells_with_indirect > 0:
            partial = round(cells_with_indirect / 8 * 0.6, 4)
            print(f"PARTIAL: Component 1 — {cells_with_indirect}/8 cells contain INDIRECT formulas "
                  f"({partial} pts). Missing: {cells_missing}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No INDIRECT formulas found in C2:C9. "
                  f"Missing: {cells_missing}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formulas match the exact expected pattern =INDIRECT(An&".B"&Bn) (0.25 points)
    # This FAILS on initial (cells are None) and should PASS on golden.
    try:
        correct_pattern = 0
        wrong_formulas = []
        for row, expected in expected_formulas.items():
            val = ws.cell(row=row, column=3).value
            if val is not None and isinstance(val, str):
                # Compare case-insensitively, removing spaces
                val_norm = val.strip().upper().replace(' ', '')
                exp_norm = expected.strip().upper().replace(' ', '')
                if val_norm == exp_norm:
                    correct_pattern += 1
                else:
                    wrong_formulas.append(f"C{row}: got {repr(val)}, expected {repr(expected)}")
            else:
                wrong_formulas.append(f"C{row}: got {repr(val)}, expected {repr(expected)}")

        if correct_pattern == 8:
            print(f"PASS: Component 2 — All 8 formulas match exact pattern =INDIRECT(An&\".B\"&Bn) (0.25 pts)")
            total_score += 0.25
        elif correct_pattern > 0:
            partial = round(correct_pattern / 8 * 0.25, 4)
            print(f"PARTIAL: Component 2 — {correct_pattern}/8 formulas match exact pattern "
                  f"({partial} pts). Issues: {wrong_formulas}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No formulas match the exact expected pattern. "
                  f"Issues: {wrong_formulas}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Summary sheet headers and A/B column data are intact (0.15 points)
    # This verifies no other cells were modified.
    # Expected headers: A1='Month', B1='Row', C1='Retrieved Value'
    # Expected A column (rows 2-9): Jan, Feb, Mar, Jan, Apr, Feb, Mar, May
    # Expected B column (rows 2-9): 5, 7, 9, 11, 5, 8, 6, 10
    # This check is designed to FAIL on initial if the task instructions say C column is empty,
    # but here we combine it: we verify the non-C cells are intact (precondition data unchanged)
    # alongside a C-column check that only passes on golden (via Component 1 already covering that).
    # To ensure this component FAILS on initial: tie it to detecting that C2:C9 have formulas
    # AND the surrounding data is intact (compound check).
    try:
        # Check headers
        header_a = ws.cell(row=1, column=1).value
        header_b = ws.cell(row=1, column=2).value
        header_c = ws.cell(row=1, column=3).value

        expected_a_vals = ['Jan', 'Feb', 'Mar', 'Jan', 'Apr', 'Feb', 'Mar', 'May']
        expected_b_vals = [5, 7, 9, 11, 5, 8, 6, 10]

        headers_ok = (
            header_a == 'Month' and
            header_b == 'Row' and
            header_c == 'Retrieved Value'
        )

        a_col_ok = all(
            ws.cell(row=r, column=1).value == expected_a_vals[r - 2]
            for r in range(2, 10)
        )
        b_col_ok = all(
            ws.cell(row=r, column=2).value == expected_b_vals[r - 2]
            for r in range(2, 10)
        )

        # This component rewards data integrity AND that C2:C9 contain formulas
        # (combining integrity check with task-change: passes only when data intact AND formulas present)
        c_has_formulas = all(
            ws.cell(row=r, column=3).value is not None and
            isinstance(ws.cell(row=r, column=3).value, str) and
            'INDIRECT' in ws.cell(row=r, column=3).value.upper()
            for r in range(2, 10)
        )

        if headers_ok and a_col_ok and b_col_ok and c_has_formulas:
            print(f"PASS: Component 3 — Headers intact, A/B columns intact, "
                  f"C2:C9 all have INDIRECT formulas (0.15 pts)")
            total_score += 0.15
        else:
            issues = []
            if not headers_ok:
                issues.append(f"Header mismatch: A1={repr(header_a)}, B1={repr(header_b)}, C1={repr(header_c)}")
            if not a_col_ok:
                actual_a = [ws.cell(row=r, column=1).value for r in range(2, 10)]
                issues.append(f"A column mismatch: got {actual_a}, expected {expected_a_vals}")
            if not b_col_ok:
                actual_b = [ws.cell(row=r, column=2).value for r in range(2, 10)]
                issues.append(f"B column mismatch: got {actual_b}, expected {expected_b_vals}")
            if not c_has_formulas:
                issues.append("Not all C2:C9 cells have INDIRECT formulas")
            print(f"FAIL: Component 3 — {'; '.join(issues)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
