"""
Reward Script: Fill in host cities for design conferences in spreadsheet
Task ID: osworld_multi_apps_conference_city_009
Domain: libreoffice_calc
Scoring:
  Component 1: SXSW rows (C2:C6) all filled with Austin city values  (0.30 pts)
  Component 2: TED Conference rows (C7:C11) all filled with Vancouver values  (0.30 pts)
  Component 3: Adobe MAX rows (C12:C16) all filled with correct city values  (0.40 pts)
Total: 1.0

Note: The task asks the agent to look up and fill in host city data from the internet.
Known ground-truth values (from task context):
  - SXSW: always Austin, TX
  - TED Conference: Vancouver, BC (2016-2020)
  - Adobe MAX 2016: San Diego, CA
  - Adobe MAX 2017: Las Vegas, NV
  - Adobe MAX 2018: Los Angeles, CA
  - Adobe MAX 2019: Los Angeles, CA
  - Adobe MAX 2020: Virtual
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_009'

# Known expected city substrings for each conference
SXSW_CITY = 'Austin'
TED_CITY = 'Vancouver'
ADOBE_MAX_CITIES = {
    2016: 'San Diego',
    2017: 'Las Vegas',
    2018: 'Los Angeles',
    2019: 'Los Angeles',
    2020: 'Virtual',
}


def _is_nonempty_city(val):
    """Return True if the cell contains a non-empty string (a city was filled in)."""
    return val is not None and str(val).strip() != ''


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the sheet
    try:
        ws = wb['DesignConferences'] if 'DesignConferences' in wb.sheetnames else wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: verify header row is intact
    try:
        if ws['A1'].value != 'Conference' or ws['B1'].value != 'Year' or ws['C1'].value != 'Host City':
            print(f"PRECONDITION FAIL: Unexpected header row: "
                  f"A1={ws['A1'].value}, B1={ws['B1'].value}, C1={ws['C1'].value}")
            print("REWARD: 0.0")
            return 0.0
        print("PRECONDITION PASS: Header row is correct")
    except Exception as e:
        print(f"PRECONDITION ERROR: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: SXSW rows (C2:C6) filled with Austin city values (0.30 points)
    # FAILS on initial (all None), PASSES on golden (all 'Austin, TX')
    try:
        sxsw_rows = [(ws[f'A{r}'].value, ws[f'B{r}'].value, ws[f'C{r}'].value) for r in range(2, 7)]
        sxsw_filled = all(_is_nonempty_city(row[2]) for row in sxsw_rows)
        sxsw_in_austin = sxsw_filled and all(
            SXSW_CITY.lower() in str(row[2]).lower() for row in sxsw_rows
        )
        if sxsw_in_austin:
            print(f"PASS: Component 1 — SXSW all 5 rows filled with Austin city (0.30 pts)")
            print(f"  Values: {[row[2] for row in sxsw_rows]}")
            total_score += 0.30
        elif sxsw_filled:
            print(f"PARTIAL: Component 1 — SXSW rows filled but city not Austin (0.15 pts)")
            print(f"  Values: {[row[2] for row in sxsw_rows]}")
            total_score += 0.15
        else:
            empty_rows = [r + 2 for r, row in enumerate(sxsw_rows) if not _is_nonempty_city(row[2])]
            print(f"FAIL: Component 1 — SXSW rows {empty_rows} have empty Host City (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: TED Conference rows (C7:C11) filled with Vancouver values (0.30 points)
    # FAILS on initial (all None), PASSES on golden (all 'Vancouver, BC')
    try:
        ted_rows = [(ws[f'A{r}'].value, ws[f'B{r}'].value, ws[f'C{r}'].value) for r in range(7, 12)]
        ted_filled = all(_is_nonempty_city(row[2]) for row in ted_rows)
        ted_in_vancouver = ted_filled and all(
            TED_CITY.lower() in str(row[2]).lower() for row in ted_rows
        )
        if ted_in_vancouver:
            print(f"PASS: Component 2 — TED Conference all 5 rows filled with Vancouver city (0.30 pts)")
            print(f"  Values: {[row[2] for row in ted_rows]}")
            total_score += 0.30
        elif ted_filled:
            print(f"PARTIAL: Component 2 — TED Conference rows filled but city not Vancouver (0.15 pts)")
            print(f"  Values: {[row[2] for row in ted_rows]}")
            total_score += 0.15
        else:
            empty_rows = [r + 7 for r, row in enumerate(ted_rows) if not _is_nonempty_city(row[2])]
            print(f"FAIL: Component 2 — TED Conference rows {empty_rows} have empty Host City (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Adobe MAX rows (C12:C16) filled with correct city values (0.40 points)
    # FAILS on initial (all None), PASSES on golden (each year has distinct correct city)
    try:
        adobe_rows = [(ws[f'A{r}'].value, ws[f'B{r}'].value, ws[f'C{r}'].value) for r in range(12, 17)]
        adobe_filled = all(_is_nonempty_city(row[2]) for row in adobe_rows)
        if adobe_filled:
            year_city_map = {row[1]: row[2] for row in adobe_rows}
            correct_count = sum(
                1 for year, expected_substr in ADOBE_MAX_CITIES.items()
                if year_city_map.get(year) is not None
                and expected_substr.lower() in str(year_city_map.get(year, '')).lower()
            )
            partial_pts = round(0.40 * correct_count / 5, 2)
            adobe_component_pts = 0.40 if correct_count == 5 else (partial_pts if correct_count >= 1 else 0.08)
            if correct_count == 5:
                print(f"PASS: Component 3 — Adobe MAX all 5 rows filled with correct cities ({adobe_component_pts} pts)")
            elif correct_count >= 1:
                print(f"PARTIAL: Component 3 — Adobe MAX {correct_count}/5 cities correct ({adobe_component_pts} pts)")
            else:
                print(f"PARTIAL: Component 3 — Adobe MAX rows filled but 0/5 match expected cities ({adobe_component_pts} pts)")
            print(f"  Values: {[(row[1], row[2]) for row in adobe_rows]}")
            if adobe_component_pts > 0:
                total_score += adobe_component_pts
        else:
            empty_rows = [r + 12 for r, row in enumerate(adobe_rows) if not _is_nonempty_city(row[2])]
            print(f"FAIL: Component 3 — Adobe MAX rows {empty_rows} have empty Host City (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
