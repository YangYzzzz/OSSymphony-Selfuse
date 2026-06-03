"""
Initial Setup: Design Conferences Host City Spreadsheet
Task ID: osworld_multi_apps_conference_city_009
Domain: libreoffice_calc

Creates DesignConferences.xlsx with Conference, Year, Host City columns.
Host City column is intentionally left blank — the agent must fill it in.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DesignConferences ---
    ws = wb.active
    ws.title = "DesignConferences"

    # Headers
    headers = ["Conference", "Year", "Host City"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, name="Calibri", size=12)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, name="Calibri", size=12, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows: 3 conferences x 5 years (2016-2020), Host City is intentionally blank
    data = [
        # SXSW: South by Southwest
        ("SXSW", 2016, ""),
        ("SXSW", 2017, ""),
        ("SXSW", 2018, ""),
        ("SXSW", 2019, ""),
        ("SXSW", 2020, ""),
        # TED Conference
        ("TED Conference", 2016, ""),
        ("TED Conference", 2017, ""),
        ("TED Conference", 2018, ""),
        ("TED Conference", 2019, ""),
        ("TED Conference", 2020, ""),
        # Adobe MAX
        ("Adobe MAX", 2016, ""),
        ("Adobe MAX", 2017, ""),
        ("Adobe MAX", 2018, ""),
        ("Adobe MAX", 2019, ""),
        ("Adobe MAX", 2020, ""),
    ]

    for row_idx, (conference, year, city) in enumerate(data, 2):
        ws.cell(row=row_idx, column=1, value=conference)
        ws.cell(row=row_idx, column=2, value=year)
        # Host City left blank (agent must fill this in)
        if city:
            ws.cell(row=row_idx, column=3, value=city)

    # Adjust column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 22

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
