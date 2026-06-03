"""
Initial Setup: Apply conditional formatting to highlight non-standard quantities
Task ID: calc_gcv_049
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Standardized_Values"

    # Headers
    ws.cell(row=1, column=1, value="Item Name")
    ws.cell(row=1, column=2, value="Quantity")

    # 19 rows of data (B2:B20)
    # Mix of standard values {10,20,30,40,50} and non-standard values
    data = [
        ("Widget Alpha", 10),
        ("Widget Beta", 15),       # non-standard
        ("Connector Pro", 30),
        ("Bracket Elite", 50),
        ("Gasket Standard", 37),   # non-standard
        ("Fastener X1", 20),
        ("Seal Ring", 42),         # non-standard
        ("Bearing Unit", 40),
        ("Spring Coil", 10),
        ("Valve Assembly", 23),    # non-standard
        ("Hinge Plate", 50),
        ("Clamp Set", 8),          # non-standard
        ("Washer Pack", 30),
        ("Bolt Kit", 55),          # non-standard
        ("Nut Assortment", 20),
        ("Pin Rod", 17),           # non-standard
        ("Anchor Bolt", 40),
        ("Rivet Box", 33),         # non-standard
        ("Screw Set", 10),
    ]

    for r, (item, qty) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=qty)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
