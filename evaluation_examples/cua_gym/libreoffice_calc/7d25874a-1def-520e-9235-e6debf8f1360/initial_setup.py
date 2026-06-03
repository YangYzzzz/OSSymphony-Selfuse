"""
Initial Setup: Lock/unlock cell protection attributes for DataEntry sheet
Task ID: calc_cop_protection_001
Domain: libreoffice_calc

Creates a DataEntry sheet with:
- Headers in A1:F1 (ID, Name, Department, Salary, Start Date, Status)
- Data rows in A2 onwards with realistic content
- ALL cells use default locked=True (the default in xlsx)
- Sheet protection is NOT enabled
"""

import openpyxl
from openpyxl.styles import Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_protection_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DataEntry ---
    ws = wb.active
    ws.title = 'DataEntry'

    # Headers in A1:F1
    headers = ['ID', 'Name', 'Department', 'Salary', 'Start Date', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        # Default: locked=True (explicitly set to make intent clear)
        cell.protection = Protection(locked=True)

    # Realistic data rows — all cells locked by default
    data = [
        [1001, 'Sarah Chen',       'Engineering',  85000, '2022-03-15', 'Active'],
        [1002, 'Marcus Johnson',   'Marketing',    72000, '2021-06-01', 'Active'],
        [1003, 'Priya Patel',      'Finance',      91000, '2020-11-20', 'Active'],
        [1004, 'Derek Williams',   'Operations',   67500, '2023-01-10', 'Active'],
        [1005, 'Aisha Okonkwo',    'HR',           63000, '2022-08-05', 'Active'],
        [1006, 'Tom Nakamura',     'Engineering',  88500, '2019-04-22', 'Active'],
        [1007, 'Lisa Fernandez',   'Sales',        74000, '2023-05-17', 'Probation'],
        [1008, 'James O\'Brien',   'Finance',      95000, '2018-09-30', 'Active'],
        [1009, 'Elena Sokolova',   'Marketing',    69000, '2021-12-03', 'Active'],
        [1010, 'Robert Osei',      'Operations',   61000, '2022-07-14', 'Active'],
        [1011, 'Hannah Schmidt',   'Engineering',  82000, '2020-02-28', 'Active'],
        [1012, 'Carlos Mendoza',   'Sales',        71500, '2023-03-09', 'Active'],
        [1013, 'Fatima Al-Hassan', 'HR',           65000, '2021-10-11', 'Active'],
        [1014, 'Liam Tremblay',    'Finance',      89000, '2019-07-16', 'Active'],
        [1015, 'Mei-Ling Wong',    'Engineering',  87000, '2022-11-25', 'Active'],
        [1016, 'Daniel Oduya',     'Operations',   64000, '2023-08-02', 'Probation'],
        [1017, 'Sophia Andreev',   'Marketing',    70500, '2020-05-19', 'Active'],
        [1018, 'Kevin Park',       'Sales',        68000, '2021-03-07', 'Active'],
        [1019, 'Amara Diallo',     'HR',           62500, '2022-09-23', 'Active'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Default: locked=True (all data cells locked by default)
            cell.protection = Protection(locked=True)

    # Sheet protection is NOT enabled — only cell attributes are set
    # (ws.protection.sheet remains False)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: DataEntry')
    print(f'  Headers: A1:F1 — {headers}')
    print(f'  Data rows: A2:A{len(data)+1} ({len(data)} rows)')
    print(f'  All cells: locked=True (default)')
    print(f'  Sheet protection: NOT enabled')


create_initial()
