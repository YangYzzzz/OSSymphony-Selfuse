"""
Reward Script: Lock header cells A1:F1 and unlock data cells A2:F100
Task ID: calc_cop_protection_001
Domain: libreoffice_calc

Scoring:
  Component 1: All 594 data cells in A2:F100 are unlocked (protection.locked=False) — 0.7 pts
  Component 2: Header cells A1:F1 remain locked AND data cells A2:F100 are all unlocked — 0.3 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_protection_001'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.

    Task: Lock header cells A1:F1 and unlock data cells A2:F100 so that
    when sheet protection is enabled, the headers are protected but the
    data rows remain editable.

    Initial state: ALL cells (headers + data) are locked (protection.locked=True)
    Golden state:  Headers A1:F1 remain locked; data A2:F100 are unlocked (locked=False)

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'DataEntry' sheet must exist
    if 'DataEntry' not in wb.sheetnames:
        print("FAIL: 'DataEntry' sheet not found — cannot evaluate")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DataEntry']

    # Component 1: All 594 data cells in A2:F100 are unlocked (0.7 points)
    # This FAILS on initial (all cells locked) → PASSES on golden (data cells unlocked)
    try:
        unlocked_count = 0
        locked_count = 0
        total_data_cells = 99 * 6  # rows 2-100, cols 1-6 = 594 cells

        for row in range(2, 101):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                if cell.protection.locked is False:
                    unlocked_count += 1
                else:
                    locked_count += 1

        if unlocked_count == total_data_cells:
            print(f"PASS: Component 1 — All {total_data_cells} data cells (A2:F100) are unlocked (0.7 pts)")
            total_score += 0.7
        else:
            print(
                f"FAIL: Component 1 — Expected all {total_data_cells} data cells unlocked, "
                f"found {unlocked_count} unlocked and {locked_count} still locked"
            )
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header cells A1:F1 are still locked AND data cells A2:F100 are all unlocked (0.3 points)
    # Combined check: passes only when BOTH conditions hold simultaneously.
    # This FAILS on initial (data cells still locked) → PASSES on golden (data unlocked, headers locked)
    try:
        header_cells_locked = all(
            ws.cell(row=1, column=col).protection.locked is True
            for col in range(1, 7)
        )
        data_cells_unlocked = all(
            ws.cell(row=row, column=col).protection.locked is False
            for row in range(2, 101)
            for col in range(1, 7)
        )

        if header_cells_locked and data_cells_unlocked:
            print(
                "PASS: Component 2 — Header cells A1:F1 are locked AND "
                "data cells A2:F100 are all unlocked (0.3 pts)"
            )
            total_score += 0.3
        elif not header_cells_locked:
            # Check which headers were accidentally unlocked
            unlocked_headers = [
                ws.cell(row=1, column=col).coordinate
                for col in range(1, 7)
                if ws.cell(row=1, column=col).protection.locked is not True
            ]
            print(
                f"FAIL: Component 2 — Header cells should remain locked, "
                f"but these were unlocked: {unlocked_headers}"
            )
        else:
            print(
                "FAIL: Component 2 — Data cells A2:F100 not all unlocked "
                "(required for combined header-lock + data-unlock check)"
            )
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
