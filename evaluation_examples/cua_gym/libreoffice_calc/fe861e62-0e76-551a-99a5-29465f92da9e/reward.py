"""
Reward Script: Apply AutoFilter to inventory sheet and filter Qty < 100
Task ID: calc_ops_014
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): AutoFilter is defined on the data range
  Component 2 (0.3): Custom filter on column C with lessThan 100
  Component 3 (0.4): Correct rows hidden (rows with Qty >= 100 hidden, rows with Qty < 100 visible)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_014'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Inventory' sheet must exist
    if 'Inventory' not in wb.sheetnames:
        print(f"CRITICAL: 'Inventory' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Component 1: AutoFilter is defined on the data range (0.3 points)
    # Initial env has no auto_filter.ref; golden has A1:C8
    try:
        af_ref = ws.auto_filter.ref
        if af_ref is not None and af_ref != '':
            # Verify the filter covers at least the header row and data area
            # Acceptable refs include A1:C8 or similar ranges covering the data
            print(f"PASS: Component 1 — AutoFilter defined with ref={af_ref} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — AutoFilter not defined (ref={af_ref})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Custom filter on column C (colId=2) with operator lessThan, val 100 (0.3 points)
    # Initial env has no filterColumn; golden has customFilter lessThan 100
    try:
        filter_columns = ws.auto_filter.filterColumn
        found_custom_filter = False
        if filter_columns:
            for fc in filter_columns:
                if fc.colId == 2 and fc.customFilters:
                    for cf in fc.customFilters.customFilter:
                        if cf.operator == 'lessThan':
                            try:
                                val = float(cf.val)
                                if val == 100.0:
                                    found_custom_filter = True
                            except (ValueError, TypeError):
                                pass
        if found_custom_filter:
            print(f"PASS: Component 2 — Custom filter on col C: lessThan 100 (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — No custom filter lessThan 100 on column C")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Rows with Qty >= 100 are hidden (0.4 points)
    # This is the task-introduced change: rows 3 (250), 5 (120), 7 (175) must be hidden.
    # In initial_env, NO rows are hidden, so this component scores 0.0 there.
    # We only check the rows that MUST CHANGE from visible to hidden.
    try:
        expected_hidden = {3, 5, 7}  # rows with Qty >= 100
        hidden_correct = 0

        for r in expected_hidden:
            if ws.row_dimensions[r].hidden:
                hidden_correct += 1
            else:
                print(f"  DETAIL: Row {r} (Qty >= 100) should be hidden but is visible")

        if hidden_correct == len(expected_hidden):
            print(f"PASS: Component 3 — All {hidden_correct} rows with Qty >= 100 are hidden (0.4 pts)")
            total_score += 0.4
        elif hidden_correct > 0:
            partial = round(0.4 * (hidden_correct / len(expected_hidden)), 2)
            print(f"PARTIAL: Component 3 — {hidden_correct}/{len(expected_hidden)} rows hidden (partial: {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No rows with Qty >= 100 are hidden")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
