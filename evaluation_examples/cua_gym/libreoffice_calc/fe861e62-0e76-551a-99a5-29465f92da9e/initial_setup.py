"""
Initial Setup: Apply AutoFilter to inventory sheet and filter Qty < 100
Task ID: calc_ops_014
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_014'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    headers = ['SKU', 'Description', 'Qty']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows - exact values from task context
    data = [
        ['P-001', 'Bearing 6205', 45],
        ['P-002', 'Seal Kit', 250],
        ['P-003', 'O-Ring Set', 80],
        ['P-004', 'Drive Belt', 120],
        ['P-005', 'Filter Element', 30],
        ['P-006', 'Gasket Set', 175],
        ['P-007', 'Impeller', 15],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10

    # NO AutoFilter, NO hidden rows - this is the pre-task state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
