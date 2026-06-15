"""
Initial Setup: Market Share Analysis - Pre-task state
Task ID: calc_sales_market_share_073
Domain: libreoffice_calc

Creates a spreadsheet with market segment data where:
- TAM and Our Revenue are populated
- Competitors Revenue is populated
- Market Penetration (D) and Status (F) columns are EMPTY (task requires filling these)
- No charts exist yet (task requires creating charts)
- No conditional formatting (task requires adding it)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_market_share_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: MarketShare ---
    ws = wb.active
    ws.title = 'MarketShare'

    # Header row styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='AAAAAA')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Set column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 20

    # Row 1: Headers
    headers = ['Segment', 'TAM', 'Our Revenue', 'Market Penetration', 'Competitors Revenue', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 32

    # Market segment data
    # Columns: Segment, TAM (raw numbers), Our Revenue, Market Penetration (EMPTY), Competitors Revenue, Status (EMPTY)
    # TAM values range $50M to $2B, Our Revenue $2M to $280M
    segments = [
        # Segment,          TAM,          Our Rev,    Competitors Rev
        ('Enterprise SaaS',   2_000_000_000,  280_000_000,  850_000_000),
        ('SMB Software',        800_000_000,   96_000_000,  310_000_000),
        ('Healthcare IT',       650_000_000,   19_500_000,  220_000_000),
        ('Financial Services',  500_000_000,   60_000_000,  185_000_000),
        ('Education Tech',      300_000_000,    9_000_000,   95_000_000),
        ('Retail Analytics',    150_000_000,   12_000_000,   52_000_000),
        ('Government & Public',  50_000_000,    2_000_000,   14_500_000),
    ]

    money_format = '$#,##0'
    data_align = Alignment(horizontal='center', vertical='center')
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    alt_fill = PatternFill(start_color='FFF5F5F5', end_color='FFF5F5F5', fill_type='solid')

    for row_idx, (seg, tam, our_rev, comp_rev) in enumerate(segments, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill(fill_type=None)

        # A: Segment name
        cell_a = ws.cell(row=row_idx, column=1, value=seg)
        cell_a.font = Font(name='Calibri', size=11)
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        cell_a.border = data_border
        if row_idx % 2 == 0:
            cell_a.fill = alt_fill

        # B: TAM
        cell_b = ws.cell(row=row_idx, column=2, value=tam)
        cell_b.number_format = money_format
        cell_b.alignment = data_align
        cell_b.border = data_border
        cell_b.font = Font(name='Calibri', size=11)
        if row_idx % 2 == 0:
            cell_b.fill = alt_fill

        # C: Our Revenue
        cell_c = ws.cell(row=row_idx, column=3, value=our_rev)
        cell_c.number_format = money_format
        cell_c.alignment = data_align
        cell_c.border = data_border
        cell_c.font = Font(name='Calibri', size=11)
        if row_idx % 2 == 0:
            cell_c.fill = alt_fill

        # D: Market Penetration — EMPTY (task requires: =C/B formula)
        cell_d = ws.cell(row=row_idx, column=4, value=None)
        cell_d.border = data_border
        if row_idx % 2 == 0:
            cell_d.fill = alt_fill

        # E: Competitors Revenue
        cell_e = ws.cell(row=row_idx, column=5, value=comp_rev)
        cell_e.number_format = money_format
        cell_e.alignment = data_align
        cell_e.border = data_border
        cell_e.font = Font(name='Calibri', size=11)
        if row_idx % 2 == 0:
            cell_e.fill = alt_fill

        # F: Status — EMPTY (task requires: =IF(D<0.05,...) formula)
        cell_f = ws.cell(row=row_idx, column=6, value=None)
        cell_f.border = data_border
        if row_idx % 2 == 0:
            cell_f.fill = alt_fill

        ws.row_dimensions[row_idx].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: MarketShare')
    print(f'  Rows: 1 header + 7 data rows (rows 2-8)')
    print(f'  Columns D (Market Penetration) and F (Status): EMPTY')
    print(f'  No charts, no conditional formatting')


create_initial()
