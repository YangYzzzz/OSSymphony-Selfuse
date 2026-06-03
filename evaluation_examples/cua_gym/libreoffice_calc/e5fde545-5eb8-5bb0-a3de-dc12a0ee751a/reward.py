"""
Reward Script: Build a market share analysis with penetration formulas,
               status flags, conditional formatting, and charts.
Task ID: calc_sales_market_share_073
Domain: libreoffice_calc
Scoring:
  Component 1: Market Penetration formulas in D2:D8 with percentage format (0.30)
  Component 2: Status IF formulas in F2:F8 flagging 'Growth Opportunity' (0.25)
  Component 3: Conditional formatting on D2:D8 (red/yellow/green rules) (0.20)
  Component 4: Pie chart present (Our Revenue vs Unaddressed Market) (0.15)
  Component 5: Bar chart present (penetration rate by segment) (0.10)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_market_share_073'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the MarketShare sheet exists
    if 'MarketShare' not in wb.sheetnames:
        print("FAIL: 'MarketShare' sheet not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['MarketShare']

    # -------------------------------------------------------------------------
    # Component 1: Market Penetration formulas D2:D8 with percentage format (0.30)
    # The task requires =C2/B2 ... =C8/B8 in column D with % format 1 decimal place
    # Initial file has D2:D8 empty; golden has formulas with 0.0% format
    # -------------------------------------------------------------------------
    try:
        penetration_formula_count = 0
        penetration_format_count = 0
        for row in range(2, 9):  # rows 2 through 8 (7 segments)
            cell = ws.cell(row=row, column=4)  # column D
            val = cell.value
            fmt = cell.number_format

            # Check that the cell has a division formula referencing C/B for the same row
            if val is not None and isinstance(val, str):
                formula_upper = val.upper().replace(' ', '')
                # Accept any formula that divides C by B for the same row
                expected = f'=C{row}/B{row}'
                if formula_upper == expected.upper():
                    penetration_formula_count += 1
                elif '/' in formula_upper and f'C{row}' in formula_upper.upper() and f'B{row}' in formula_upper.upper():
                    penetration_formula_count += 1

            # Check percentage number format (should contain % and be for 1 decimal)
            if fmt and '%' in fmt:
                penetration_format_count += 1

        if penetration_formula_count == 7:
            print(f"PASS: Component 1a — All 7 Market Penetration formulas present in D2:D8")
            formula_score = 0.20
        elif penetration_formula_count >= 4:
            print(f"PARTIAL: Component 1a — {penetration_formula_count}/7 penetration formulas found in D2:D8")
            formula_score = 0.10
        else:
            print(f"FAIL: Component 1a — Only {penetration_formula_count}/7 penetration formulas in D2:D8")
            formula_score = 0.0

        if penetration_format_count == 7:
            print(f"PASS: Component 1b — All 7 cells in D2:D8 have percentage number format")
            format_score = 0.10
        elif penetration_format_count >= 4:
            print(f"PARTIAL: Component 1b — {penetration_format_count}/7 cells with percentage format in D2:D8")
            format_score = 0.05
        else:
            print(f"FAIL: Component 1b — Only {penetration_format_count}/7 cells with % format in D2:D8")
            format_score = 0.0

        total_score += formula_score + format_score

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Status IF formulas in F2:F8 (0.25)
    # The task requires =IF(D2<0.05,"Growth Opportunity","Established") pattern
    # Initial file has F2:F8 empty; golden has IF formulas
    # -------------------------------------------------------------------------
    try:
        status_formula_count = 0
        for row in range(2, 9):  # rows 2 through 8
            cell = ws.cell(row=row, column=6)  # column F
            val = cell.value

            if val is not None and isinstance(val, str):
                formula_upper = val.upper().replace(' ', '')
                # Must be an IF formula referencing D{row}<0.05 and contain "GROWTHOPPORTUNITY"
                d_ref = f'D{row}'
                if (formula_upper.startswith('=IF(') and
                        d_ref.upper() in formula_upper and
                        '0.05' in formula_upper and
                        'GROWTHOPPORTUNITY' in formula_upper.upper()):
                    status_formula_count += 1

        if status_formula_count == 7:
            print(f"PASS: Component 2 — All 7 Status IF formulas present in F2:F8 with 'Growth Opportunity' flag")
            total_score += 0.25
        elif status_formula_count >= 4:
            print(f"PARTIAL: Component 2 — {status_formula_count}/7 Status IF formulas in F2:F8")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — Only {status_formula_count}/7 status formulas in F2:F8 (expected IF with Growth Opportunity)")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Conditional formatting on D2:D8 with 3 color rules (0.20)
    # The task requires: red if <0.05, yellow if <0.15, green if >=0.15
    # Initial file has no conditional formatting; golden has 3 rules on D2:D8
    # -------------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        cf_ranges = list(cf_rules)

        # Look for conditional formatting that covers D column (rows 2-8)
        d_col_cf_found = False
        rule_count = 0
        has_red_lt_5pct = False
        has_yellow_rule = False
        has_green_gte_15pct = False

        for cf_range in cf_ranges:
            range_str = str(cf_range)
            # Check if this CF covers D2:D8 range (or similar)
            if 'D' in range_str:
                d_col_cf_found = True
                for rule in cf_range.rules:
                    rule_count += 1
                    rule_formula = getattr(rule, 'formula', None)
                    rule_op = getattr(rule, 'operator', None)

                    # Check for red rule: cellIs < 0.05
                    if (rule_op == 'lessThan' and rule_formula and
                            '0.05' in str(rule_formula)):
                        has_red_lt_5pct = True

                    # Check for yellow rule: between 0.05 and 0.15
                    if (rule_op == 'between' and rule_formula and
                            '0.05' in str(rule_formula) and '0.15' in str(rule_formula)):
                        has_yellow_rule = True

                    # Check for green rule: >= 0.15
                    if (rule_op == 'greaterThanOrEqual' and rule_formula and
                            '0.15' in str(rule_formula)):
                        has_green_gte_15pct = True

        if d_col_cf_found and has_red_lt_5pct and has_yellow_rule and has_green_gte_15pct:
            print("PASS: Component 3 — Conditional formatting on D column with all 3 color rules (red/yellow/green)")
            total_score += 0.20
        elif d_col_cf_found and rule_count >= 2:
            print(f"PARTIAL: Component 3 — Conditional formatting on D column found with {rule_count} rules (missing some: red={has_red_lt_5pct}, yellow={has_yellow_rule}, green={has_green_gte_15pct})")
            total_score += 0.10
        elif d_col_cf_found:
            print(f"PARTIAL: Component 3 — Conditional formatting found on D column but only {rule_count} rule(s)")
            total_score += 0.05
        else:
            print("FAIL: Component 3 — No conditional formatting found on column D (D2:D8)")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Pie chart present showing Our Revenue vs Unaddressed Market (0.15)
    # Initial file has no charts; golden has a PieChart
    # -------------------------------------------------------------------------
    try:
        all_charts = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            charts = getattr(sheet, '_charts', [])
            all_charts.extend(charts)

        from openpyxl.chart import PieChart

        pie_charts = [c for c in all_charts if isinstance(c, PieChart)]

        if len(pie_charts) >= 1:
            print(f"PASS: Component 4 — Pie chart found ({len(pie_charts)} pie chart(s) in workbook)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — No pie chart found in workbook (found chart types: {[type(c).__name__ for c in all_charts]})")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Bar chart present showing penetration rate by segment (0.10)
    # Initial file has no charts; golden has a BarChart
    # -------------------------------------------------------------------------
    try:
        from openpyxl.chart import BarChart

        bar_charts = [c for c in all_charts if isinstance(c, BarChart)]

        if len(bar_charts) >= 1:
            print(f"PASS: Component 5 — Bar chart found ({len(bar_charts)} bar chart(s) in workbook)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No bar chart found in workbook")

    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
