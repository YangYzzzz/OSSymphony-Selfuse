"""
Initial Setup: Clear formatting from A1:G1 header row (pre-task state)
Task ID: calc_cop_clear_002
Domain: libreoffice_calc

This script creates the initial spreadsheet with a styled header row
(bold, 14pt, dark blue background, white text, all borders) that the
agent is expected to clear formatting from while keeping text values.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_clear_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: StyledSheet ---
    ws = wb.active
    ws.title = 'StyledSheet'

    # Header row A1:G1 with styling
    headers = ['Name', 'Dept', 'Role', 'Salary', 'Start Date', 'Status', 'Notes']

    # Define styles for header row
    dark_blue_fill = PatternFill(start_color='FF003366', end_color='FF003366', fill_type='solid')
    white_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    thin = Side(style='thin', color='000000')
    all_borders = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = dark_blue_fill
        cell.border = all_borders
        cell.alignment = center_align

    # Set header row height
    ws.row_dimensions[1].height = 22

    # Realistic employee data rows (no special formatting)
    data = [
        ['Sarah Chen',      'Engineering',  'Senior Engineer',   92000,  '2021-03-10', 'Active',   'Team lead candidate'],
        ['Marcus Johnson',  'Marketing',    'Marketing Analyst', 68500,  '2022-07-01', 'Active',   'Q3 campaign lead'],
        ['Priya Patel',     'Finance',      'Financial Analyst', 74000,  '2020-11-15', 'Active',   'CPA certified'],
        ['James Whitfield', 'Engineering',  'Junior Engineer',   58000,  '2023-01-20', 'Active',   'New hire'],
        ['Liu Wei',         'Sales',        'Account Manager',   81000,  '2019-06-05', 'Active',   'Top performer 2024'],
        ['Anika Osei',      'HR',           'HR Specialist',     62000,  '2021-09-30', 'Active',   'Benefits coordinator'],
        ['Carlos Rivera',   'Operations',   'Ops Manager',       87500,  '2018-04-22', 'Active',   'Certified PMP'],
        ['Emma Larson',     'Engineering',  'QA Engineer',       71000,  '2022-02-14', 'On Leave', 'Parental leave'],
        ['David Okonkwo',   'Marketing',    'Brand Strategist',  76000,  '2020-08-01', 'Active',   'Social media expert'],
        ['Fatima Al-Rashid','Finance',      'Controller',        95000,  '2017-12-11', 'Active',   'Department head'],
        ['Tom Bergstrom',   'Sales',        'Sales Rep',         54000,  '2023-05-03', 'Probation','First 90 days'],
        ['Nina Castillo',   'Operations',   'Logistics Coord.',  60500,  '2022-10-17', 'Active',   'Warehouse oversight'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    col_widths = [20, 14, 22, 12, 14, 12, 25]
    for i, width in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: StyledSheet')
    print(f'  Header row A1:G1: bold, 14pt, dark blue bg (#003366), white text, all borders')
    print(f'  Data rows: 12 employee records (rows 2-13)')


create_initial()
