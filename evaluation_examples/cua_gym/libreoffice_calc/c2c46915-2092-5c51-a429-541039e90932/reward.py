"""
Reward Script: Clear only the formatting from the range A1:G1 while keeping header text values intact.
Task ID: calc_cop_clear_002
Domain: libreoffice_calc
Scoring:
  Component 1: Font bold removed from all A1:G1 cells       — 0.30 pts
  Component 2: Background fill removed from all A1:G1 cells  — 0.30 pts
  Component 3: All borders removed from all A1:G1 cells       — 0.20 pts
  Component 4: Font size and font color reset in A1:G1        — 0.20 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_clear_002'
SHEET_NAME = 'StyledSheet'
HEADER_COLS = 7  # A through G

# Expected header text values (used as a precondition gate only)
EXPECTED_HEADERS = ['Name', 'Dept', 'Role', 'Salary', 'Start Date', 'Status', 'Notes']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires clearing ALL formatting from A1:G1 while preserving text values.
    Initial state: bold=True, size=14, white font color, dark blue background (FF003366), thin borders.
    Golden state: no bold, no size, no font color, no fill, no borders.
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Precondition gate: header text values must be intact
    # If text values are missing or changed, the task was performed incorrectly — do not award any score.
    for col in range(1, HEADER_COLS + 1):
        cell = ws.cell(row=1, column=col)
        actual = cell.value
        expected = EXPECTED_HEADERS[col - 1]
        if actual != expected:
            print(f"CRITICAL: Header text in column {col} changed: expected {repr(expected)}, found {repr(actual)}")
            print("REWARD: 0.0")
            return 0.0
    print("GATE PASS: All A1:G1 header text values are intact.")

    # Component 1: Bold removed from all A1:G1 cells (0.30 points)
    # Initial state: bold=True for all 7 cells
    # Golden state: bold=False (or None) for all 7 cells
    try:
        bold_cleared = True
        bold_failures = []
        for col in range(1, HEADER_COLS + 1):
            cell = ws.cell(row=1, column=col)
            if cell.font.bold:
                bold_cleared = False
                from openpyxl.utils import get_column_letter
                bold_failures.append(f"{get_column_letter(col)}1")

        if bold_cleared:
            print(f"PASS: Component 1 — Bold formatting removed from all A1:G1 cells (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Bold still present in cells: {bold_failures}")
    except Exception as e:
        print(f"ERROR: Component 1 (bold check) — {e}")

    # Component 2: Background fill removed from all A1:G1 cells (0.30 points)
    # Initial state: fill_type='solid', fgColor='FF003366' (dark blue background)
    # Golden state: fill_type=None (no fill)
    try:
        fill_cleared = True
        fill_failures = []
        for col in range(1, HEADER_COLS + 1):
            cell = ws.cell(row=1, column=col)
            # A cell is considered to have a fill if patternType is 'solid'
            if cell.fill.patternType == 'solid':
                fill_cleared = False
                from openpyxl.utils import get_column_letter
                try:
                    color_info = cell.fill.fgColor.rgb
                except Exception:
                    color_info = 'unknown'
                fill_failures.append(f"{get_column_letter(col)}1 (color={color_info})")

        if fill_cleared:
            print(f"PASS: Component 2 — Background fill removed from all A1:G1 cells (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — Solid fill still present in cells: {fill_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 (fill check) — {e}")

    # Component 3: All borders removed from all A1:G1 cells (0.20 points)
    # Initial state: thin borders on all 4 sides of every cell
    # Golden state: no borders (all sides are None)
    try:
        borders_cleared = True
        border_failures = []
        for col in range(1, HEADER_COLS + 1):
            cell = ws.cell(row=1, column=col)
            from openpyxl.utils import get_column_letter
            cell_ref = f"{get_column_letter(col)}1"
            for side_name in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, side_name)
                if side and side.style is not None:
                    borders_cleared = False
                    border_failures.append(f"{cell_ref} {side_name}={side.style}")

        if borders_cleared:
            print(f"PASS: Component 3 — All borders removed from A1:G1 cells (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Borders still present: {border_failures[:5]}")  # show first 5 failures
    except Exception as e:
        print(f"ERROR: Component 3 (border check) — {e}")

    # Component 4: Font size and font color reset in A1:G1 (0.20 points)
    # Initial state: size=14.0, color='00FFFFFF' (white)
    # Golden state: size=None (default), no explicit color (font.color is None or has no rgb)
    # Award partial: 0.10 for size cleared, 0.10 for color cleared
    try:
        size_cleared = True
        color_cleared = True
        size_failures = []
        color_failures = []
        from openpyxl.utils import get_column_letter

        for col in range(1, HEADER_COLS + 1):
            cell = ws.cell(row=1, column=col)
            cell_ref = f"{get_column_letter(col)}1"

            # Size check: initial was 14.0; golden should be None (default)
            if cell.font.size is not None and cell.font.size != 11:
                # Allow size=11 as a valid default (LibreOffice default font size)
                size_cleared = False
                size_failures.append(f"{cell_ref}={cell.font.size}")

            # Color check: initial was white (00FFFFFF); golden should have no explicit color
            try:
                font_color = cell.font.color
                if font_color is not None and font_color.rgb is not None:
                    # Any explicit non-black color is a failure (white=00FFFFFF was the initial)
                    rgb = font_color.rgb
                    # Accept black (FF000000, 00000000) and None as cleared
                    if rgb not in ('00000000', 'FF000000', None):
                        color_cleared = False
                        color_failures.append(f"{cell_ref}={rgb}")
            except (AttributeError, TypeError):
                # font.color or font.color.rgb is None — color is cleared (good)
                pass

        comp4_score = 0.0
        if size_cleared:
            print(f"PASS: Component 4a — Font size reset to default in all A1:G1 cells (0.10 pts)")
            comp4_score += 0.10
        else:
            print(f"FAIL: Component 4a — Font size not reset: {size_failures}")

        if color_cleared:
            print(f"PASS: Component 4b — Font color reset to default in all A1:G1 cells (0.10 pts)")
            comp4_score += 0.10
        else:
            print(f"FAIL: Component 4b — Font color still explicit: {color_failures}")

        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 (font size/color check) — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
