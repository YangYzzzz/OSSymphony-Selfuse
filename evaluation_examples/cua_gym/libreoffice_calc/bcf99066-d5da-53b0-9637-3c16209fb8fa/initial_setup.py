"""
Initial Setup: Create SalesDB spreadsheet with 240 rows of sales data
Task ID: calc_pivot_040
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SalesDB'

    # Headers
    headers = ['ID', 'Region', 'Category', 'Sales', 'Profit']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    regions = ['North', 'South', 'East', 'West']
    categories = ['Electronics', 'Apparel', 'Home', 'Sports']

    # Target subtotals per region:
    # North=78000, East=72000, South=65000, West=55000, total=270000
    # 60 rows per region, 15 rows per (region, category)
    region_targets = {
        'North': 78000,
        'South': 65000,
        'East': 72000,
        'West': 55000,
    }

    rows_data = []

    for region in regions:
        target = region_targets[region]
        # Generate 60 random base values
        raw_values = [random.randint(200, 600) for _ in range(60)]
        raw_sum = sum(raw_values)

        # Scale to hit exact target
        scaled = []
        running_sum = 0
        for i in range(60):
            if i < 59:
                sv = round(raw_values[i] * target / raw_sum)
                scaled.append(sv)
                running_sum += sv
            else:
                scaled.append(target - running_sum)

        # Assign categories round-robin: each category gets 15 rows
        for i in range(60):
            cat = categories[i % 4]
            sales = scaled[i]
            profit = round(sales * random.uniform(0.15, 0.35))
            rows_data.append((region, cat, sales, profit))

    # Shuffle to make data look realistic (not grouped by region)
    random.shuffle(rows_data)

    # Write to sheet with sequential IDs
    for idx, (region, category, sales, profit) in enumerate(rows_data):
        r = idx + 2
        ws.cell(row=r, column=1, value=idx + 1)
        ws.cell(row=r, column=2, value=region)
        ws.cell(row=r, column=3, value=category)
        ws.cell(row=r, column=4, value=sales)
        ws.cell(row=r, column=5, value=profit)

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify region totals
    region_sums = {}
    for row in ws.iter_rows(min_row=2, max_row=241, min_col=2, max_col=4):
        reg = row[0].value
        sal = row[2].value
        region_sums[reg] = region_sums.get(reg, 0) + sal
    print(f'Region sums: {region_sums}')
    print(f'Grand total: {sum(region_sums.values())}')

    # Launch LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
