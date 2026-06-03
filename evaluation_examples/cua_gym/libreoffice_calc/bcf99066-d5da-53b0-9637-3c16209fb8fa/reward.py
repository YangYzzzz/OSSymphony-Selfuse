"""
Reward Script: Create pivot table with Region/Category nested rows, SUM of Sales, sorted descending
Task ID: calc_pivot_040
Domain: libreoffice_calc
Scoring:
  Component 1: PivotTable sheet exists (0.10)
  Component 2: Correct column headers (0.10)
  Component 3: Region subtotals correct (0.30)
  Component 4: Grand total correct (0.10)
  Component 5: Regions sorted by sales descending (0.20)
  Component 6: Categories nested under regions with correct values (0.20)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_040'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.10 points)
    # This sheet does NOT exist in initial_env, only in golden_env
    try:
        if 'PivotTable' in wb.sheetnames:
            print(f"PASS: Component 1 — PivotTable sheet exists (0.10 pts)")
            total_score += 0.10
        else:
            # Check for common alternative names
            pivot_sheet = None
            for sn in wb.sheetnames:
                if 'pivot' in sn.lower():
                    pivot_sheet = sn
                    break
            if pivot_sheet:
                print(f"PARTIAL: Component 1 — Found pivot sheet named '{pivot_sheet}' instead of 'PivotTable' (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 1 — No PivotTable sheet found. Sheets: {wb.sheetnames}")
                print("REWARD: 0.0")
                return 0.0
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the pivot sheet (use PivotTable or fallback to any pivot-named sheet)
    pivot_name = 'PivotTable'
    if pivot_name not in wb.sheetnames:
        for sn in wb.sheetnames:
            if 'pivot' in sn.lower():
                pivot_name = sn
                break
    ws = wb[pivot_name]

    # Read all data from pivot sheet into structured form
    all_data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        all_data.append(row)

    if len(all_data) < 2:
        print(f"FAIL: PivotTable sheet has too few rows ({len(all_data)})")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Correct column headers (0.10 points)
    # Headers should include Region, Category, and a Sales-related column
    try:
        headers = [str(h).strip().lower() if h else '' for h in all_data[0]]
        has_region = any('region' in h for h in headers)
        has_category = any('category' in h or 'cat' in h for h in headers)
        has_sales = any('sales' in h or 'sum' in h for h in headers)

        if has_region and has_category and has_sales:
            print(f"PASS: Component 2 — Headers correct: {all_data[0]} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Missing headers. Found: {all_data[0]}. Need Region, Category, Sales/Sum")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Parse the pivot table structure
    # Expected: rows with Region+Category for detail, Region Total for subtotals, Grand Total at end
    # Identify the sales column index (0-based in all_data tuples)
    sales_col = None
    for i, h in enumerate(all_data[0]):
        if h and ('sales' in str(h).lower() or 'sum' in str(h).lower()):
            sales_col = i
            break
    if sales_col is None:
        sales_col = len(all_data[0]) - 1  # fallback: last column

    region_col = 0  # Region is typically first column
    cat_col = 1     # Category is typically second column

    # Extract region subtotals and grand total from the data
    region_subtotals = {}
    grand_total_value = None
    region_order = []
    region_categories = {}  # region -> [(category, sales_value), ...]

    current_region = None
    for row_data in all_data[1:]:  # skip header
        a_val = str(row_data[region_col]).strip() if row_data[region_col] else ''
        b_val = str(row_data[cat_col]).strip() if row_data[cat_col] else ''
        c_val = row_data[sales_col]

        # Detect Grand Total row
        if 'grand' in a_val.lower() and 'total' in a_val.lower():
            try:
                grand_total_value = float(c_val) if c_val else None
            except (ValueError, TypeError):
                grand_total_value = None
            continue

        # Detect Region Total row (e.g. "North Total")
        if 'total' in a_val.lower():
            region_name = a_val.lower().replace('total', '').strip().title()
            try:
                region_subtotals[region_name] = float(c_val) if c_val else None
            except (ValueError, TypeError):
                region_subtotals[region_name] = None
            if region_name not in region_order:
                region_order.append(region_name)
            continue

        # Detail row: Region + Category
        if a_val and a_val.lower() not in ('none', ''):
            current_region = a_val.strip()
            if current_region not in region_order:
                region_order.append(current_region)
            if current_region not in region_categories:
                region_categories[current_region] = []

        if current_region and b_val and b_val.lower() not in ('none', ''):
            try:
                cat_sales = float(c_val) if c_val else 0
            except (ValueError, TypeError):
                cat_sales = 0
            region_categories[current_region].append((b_val, cat_sales))

    print(f"\nDEBUG: Region subtotals found: {region_subtotals}")
    print(f"DEBUG: Grand total: {grand_total_value}")
    print(f"DEBUG: Region order: {region_order}")
    print(f"DEBUG: Region categories: {region_categories}")

    # Component 3: Region subtotals correct (0.30 points)
    # Expected: North=78000, East=72000, South=65000, West=55000
    expected_subtotals = {
        'North': 78000,
        'East': 72000,
        'South': 65000,
        'West': 55000,
    }
    try:
        subtotal_matches = 0
        for region, expected_val in expected_subtotals.items():
            actual = region_subtotals.get(region)
            if actual is not None and abs(actual - expected_val) < 1.0:
                subtotal_matches += 1
                print(f"  PASS: {region} subtotal = {actual} (expected {expected_val})")
            else:
                print(f"  FAIL: {region} subtotal = {actual} (expected {expected_val})")

        if subtotal_matches == 4:
            print(f"PASS: Component 3 — All 4 region subtotals correct (0.30 pts)")
            total_score += 0.30
        elif subtotal_matches >= 2:
            partial = round(0.30 * subtotal_matches / 4, 2)
            print(f"PARTIAL: Component 3 — {subtotal_matches}/4 subtotals correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {subtotal_matches}/4 subtotals correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand total correct (0.10 points)
    # Expected: 270000
    try:
        if grand_total_value is not None and abs(grand_total_value - 270000) < 1.0:
            print(f"PASS: Component 4 — Grand total = {grand_total_value} (expected 270000) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Grand total = {grand_total_value} (expected 270000)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Regions sorted by sales descending (0.20 points)
    # Expected order: North (78000) > East (72000) > South (65000) > West (55000)
    try:
        expected_region_order = ['North', 'East', 'South', 'West']
        # Normalize region_order to title case for comparison
        normalized_order = [r.strip().title() for r in region_order if r.strip().title() in expected_region_order]

        if normalized_order == expected_region_order:
            print(f"PASS: Component 5 — Regions sorted descending by sales: {normalized_order} (0.20 pts)")
            total_score += 0.20
        else:
            # Check if all 4 regions present but wrong order
            if set(normalized_order) == set(expected_region_order):
                print(f"FAIL: Component 5 — Regions present but wrong order: {normalized_order} (expected {expected_region_order})")
            else:
                print(f"FAIL: Component 5 — Region order: {normalized_order} (expected {expected_region_order})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Categories nested under each region with correct values (0.20 points)
    # Check that each region has 4 categories and they're sorted by sales descending within each region
    try:
        cat_score = 0.0
        regions_with_cats = 0
        regions_sorted_desc = 0

        for region in expected_subtotals.keys():
            cats = region_categories.get(region, [])
            if len(cats) >= 4:
                regions_with_cats += 1
                # Check descending sort within region
                sales_vals = [s for _, s in cats]
                if sales_vals == sorted(sales_vals, reverse=True):
                    regions_sorted_desc += 1

        # 0.10 for having categories nested under all regions
        if regions_with_cats == 4:
            cat_score += 0.10
            print(f"  PASS: All 4 regions have category details")
        elif regions_with_cats > 0:
            partial = round(0.10 * regions_with_cats / 4, 2)
            cat_score += partial
            print(f"  PARTIAL: {regions_with_cats}/4 regions have category details")
        else:
            print(f"  FAIL: No regions have category details")

        # 0.10 for categories sorted descending within each region
        if regions_sorted_desc == 4:
            cat_score += 0.10
            print(f"  PASS: Categories sorted descending in all regions")
        elif regions_sorted_desc > 0:
            partial = round(0.10 * regions_sorted_desc / 4, 2)
            cat_score += partial
            print(f"  PARTIAL: {regions_sorted_desc}/4 regions have categories sorted descending")
        else:
            print(f"  FAIL: Categories not sorted descending in any region")

        if cat_score > 0:
            print(f"PASS: Component 6 — Category nesting and sorting ({cat_score} pts)")
            total_score += cat_score
        else:
            print(f"FAIL: Component 6 — No category nesting found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
