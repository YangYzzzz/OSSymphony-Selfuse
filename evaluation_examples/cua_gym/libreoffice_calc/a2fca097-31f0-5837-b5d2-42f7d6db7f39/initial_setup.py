"""
Initial Setup: Define named range and data validation for employee list
Task ID: calc_mcp_049
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Employees ---
    ws1 = wb.active
    ws1.title = 'Employees'
    ws1['A1'] = 'Name'
    ws1['A1'].font = openpyxl.styles.Font(bold=True)

    employee_names = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
        'Yuki Tanaka', 'Elena Rodriguez', 'David Kim', 'Fatima Al-Rashid',
        'Thomas Mueller', 'Amara Okafor', 'Liam Fitzgerald', 'Mei-Lin Chang',
        'Carlos Gutierrez', 'Anya Petrov', 'Raj Sharma', 'Isabella Rossi',
        'Kwame Asante', 'Sofia Andersen', 'Viktor Novak', 'Aisha Mohammed',
        'Benjamin Foster', 'Hana Watanabe', 'Lucas Silva', 'Nadia Kowalski',
        'Omar Hassan', 'Chloe Dupont', 'Ravi Krishnamurthy', 'Emily Watson',
        'Andrei Volkov', 'Zara Nkemelu', 'Michael Thompson', 'Lin Xiaoming',
        'Jessica Martinez', 'Patrick O\'Sullivan', 'Sakura Mori',
        'Alejandro Reyes', 'Freya Johansson', 'Deepak Sundaram',
        'Catherine Lambert', 'Tariq Mansour', 'Rachel Green', 'Hiroshi Nakamura',
        'Claudia Bianchi', 'Samuel Oduya', 'Ingrid Bergstrom', 'Wei Zhang',
        'Natasha Ivanova', 'Felipe Cardoso', 'Grace Mensah', 'Jakob Hoffman',
        'Sunita Desai', 'Pierre Moreau', 'Yolanda Cruz', 'Aleksei Popov',
        'Margaret O\'Neill', 'Kenji Sato', 'Lucia Fernandez', 'Oluwole Adeyemi',
        'Emma Lindqvist', 'Hassan Youssef', 'Dmitri Sokolov', 'Ana Pereira',
        'Brian Mitchell', 'Chun Hei Wong', 'Diana Vasilescu', 'Emeka Obi',
        'Fiona MacLeod', 'Giovanni Conti', 'Helen Papadopoulos', 'Ibrahim Syed',
        'Jana Horvat', 'Kevin O\'Malley', 'Layla Khoury', 'Mateo Alvarez',
        'Nina Johansson', 'Oscar Lundgren', 'Paloma Vega', 'Quentin Fabre',
        'Rosa Colombo', 'Stefan Becker', 'Tanya Reddy', 'Ulrich Braun',
        'Valentina Soares', 'William Park', 'Xia Li', 'Yasmin Akhtar',
        'Zachary Burns', 'Arjun Nair', 'Beatrice Fontaine', 'Chidi Eze',
        'Daniela Morales', 'Erik Haugen', 'Gabriela Rojas', 'Hugo Martins',
        'Irene Papadakis', 'Jun Takahashi', 'Katarina Novotna', 'Lorenzo Ricci',
        'Mina Saeed', 'Nicholas Grant', 'Olga Kuznetsova',
    ]

    for i, name in enumerate(employee_names, 2):
        ws1.cell(row=i, column=1, value=name)

    # Set column width for readability
    ws1.column_dimensions['A'].width = 28

    # --- Sheet 2: Validation ---
    ws2 = wb.create_sheet('Validation')
    ws2['A1'] = 'Task ID'
    ws2['A1'].font = openpyxl.styles.Font(bold=True)
    ws2['B1'] = 'Assigned To'
    ws2['B1'].font = openpyxl.styles.Font(bold=True)

    # Add some task IDs in column A for context
    tasks = [
        'PROJ-101', 'PROJ-102', 'PROJ-103', 'PROJ-104', 'PROJ-105',
        'PROJ-106', 'PROJ-107', 'PROJ-108', 'PROJ-109', 'PROJ-110',
        'PROJ-111', 'PROJ-112', 'PROJ-113', 'PROJ-114', 'PROJ-115',
        'PROJ-116', 'PROJ-117', 'PROJ-118', 'PROJ-119',
    ]
    for i, task in enumerate(tasks, 2):
        ws2.cell(row=i, column=1, value=task)

    # B2:B20 are intentionally left empty - no data validation yet
    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 28

    # NO named ranges defined
    # NO data validation applied

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
