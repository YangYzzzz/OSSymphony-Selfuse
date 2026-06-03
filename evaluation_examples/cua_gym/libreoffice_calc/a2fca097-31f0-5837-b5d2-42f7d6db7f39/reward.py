"""
Reward Script: Define named range 'EmployeeList' and use it for data validation dropdown
Task ID: calc_mcp_049
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Named range 'EmployeeList' exists covering Employees!$A$2:$A$500
  Component 2 (0.5): Data validation dropdown on Validation!B2:B20 sourced from EmployeeList
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_049'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both sheets must exist
    if 'Employees' not in wb.sheetnames or 'Validation' not in wb.sheetnames:
        print(f"FAIL: Required sheets missing. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Named range 'EmployeeList' exists and covers Employees!$A$2:$A$500 (0.5 points)
    try:
        found_named_range = False
        correct_range = False

        # Check if 'EmployeeList' defined name exists
        if 'EmployeeList' in wb.defined_names:
            found_named_range = True
            dn = wb.defined_names['EmployeeList']
            destinations = list(dn.destinations)
            # destinations is list of (sheet_title, cell_range) tuples
            # Expected: [('Employees', '$A$2:$A$500')]
            for sheet_title, cell_range in destinations:
                if sheet_title == 'Employees' and cell_range == '$A$2:$A$500':
                    correct_range = True
                    break

            if correct_range:
                print(f"PASS: Component 1 — Named range 'EmployeeList' = Employees!$A$2:$A$500 (0.5 pts)")
                total_score += 0.5
            else:
                # Partial: named range exists but wrong range
                print(f"FAIL: Component 1 — 'EmployeeList' exists but destinations={destinations}, expected Employees!$A$2:$A$500")
        else:
            print(f"FAIL: Component 1 — No defined name 'EmployeeList' found. Available: {list(wb.defined_names.keys())}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data validation on Validation!B2:B20 is list type sourced from EmployeeList (0.5 points)
    try:
        ws_val = wb['Validation']
        validations = ws_val.data_validations.dataValidation

        if len(validations) == 0:
            print(f"FAIL: Component 2 — No data validations found on 'Validation' sheet")
        else:
            found_correct_dv = False
            for dv in validations:
                # Check type is 'list'
                if dv.type != 'list':
                    continue

                # Check formula references EmployeeList
                formula = str(dv.formula1) if dv.formula1 else ''
                # The formula could be '=EmployeeList' or 'EmployeeList'
                if 'EmployeeList' not in formula:
                    continue

                # Check sqref covers B2:B20
                sqref_str = str(dv.sqref)
                if 'B2:B20' in sqref_str:
                    found_correct_dv = True
                    print(f"PASS: Component 2 — Data validation list with formula='{formula}' on {sqref_str} (0.5 pts)")
                    total_score += 0.5
                    break

            if not found_correct_dv:
                # Print what we found for debugging
                for dv in validations:
                    print(f"  Found DV: type={dv.type}, formula1={dv.formula1}, sqref={dv.sqref}")
                print(f"FAIL: Component 2 — No matching data validation (list type, EmployeeList source, B2:B20 range)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
