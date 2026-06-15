"""
Initial Setup: Create a spreadsheet with 10 columns and 500 rows of business data.
Task ID: calc_nrv_033
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers for 10 columns (A-J)
    headers = [
        "Employee ID", "Full Name", "Department", "Position",
        "Annual Salary", "Hire Date", "Region", "Manager",
        "Performance Score", "Bonus Amount"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data pools
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei",
        "Carlos", "Aisha", "Thomas", "Olivia", "Raj", "Fatima", "Liam",
        "Yuki", "Andre", "Sophia", "Wei", "Amara", "Noah", "Isabella",
        "Kenji", "Zara", "Ethan", "Leila", "Ryan", "Nia", "Oscar",
        "Hannah", "Viktor"
    ]
    last_names = [
        "Chen", "Johnson", "Petrov", "Williams", "Sharma", "Kim",
        "Rodriguez", "Anderson", "Okafor", "Martinez", "Taylor",
        "Nakamura", "Hassan", "Brown", "Tanaka", "Davis", "Singh",
        "Garcia", "Wilson", "Lee", "Moore", "Patel", "Clark",
        "Nguyen", "Hall", "Lopez", "Scott", "Adams", "Baker", "Rivera"
    ]
    departments = [
        "Engineering", "Marketing", "Finance", "Human Resources",
        "Operations", "Sales", "Research", "Customer Support",
        "Legal", "Product Management"
    ]
    positions = [
        "Analyst", "Senior Analyst", "Manager", "Director",
        "Specialist", "Coordinator", "Lead", "Associate",
        "Consultant", "Engineer"
    ]
    regions = [
        "North America", "Europe", "Asia Pacific", "Latin America",
        "Middle East", "Africa"
    ]
    managers = [
        "Victoria Palmer", "Robert Chang", "Diana Foster",
        "Michael Torres", "Jennifer Wu", "Stephen Blake",
        "Maria Santos", "Alexander Reed", "Lisa Hoffman", "Kevin Pham"
    ]

    # Generate 499 data rows (rows 2-500)
    for r in range(2, 501):
        emp_id = f"EMP-{r - 1:04d}"
        full_name = f"{random.choice(first_names)} {random.choice(last_names)}"
        dept = random.choice(departments)
        pos = random.choice(positions)
        salary = random.randint(45000, 180000)
        year = random.randint(2015, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        hire_date = f"{year}-{month:02d}-{day:02d}"
        region = random.choice(regions)
        manager = random.choice(managers)
        perf_score = round(random.uniform(1.0, 5.0), 1)
        bonus = round(salary * random.uniform(0.02, 0.15), 2)

        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=full_name)
        ws.cell(row=r, column=3, value=dept)
        ws.cell(row=r, column=4, value=pos)
        ws.cell(row=r, column=5, value=salary)
        ws.cell(row=r, column=6, value=hire_date)
        ws.cell(row=r, column=7, value=region)
        ws.cell(row=r, column=8, value=manager)
        ws.cell(row=r, column=9, value=perf_score)
        ws.cell(row=r, column=10, value=bonus)

    # NO named ranges, NO print ranges in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
