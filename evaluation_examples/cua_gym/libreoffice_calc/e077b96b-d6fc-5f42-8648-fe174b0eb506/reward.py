"""
Reward Script: Create named range 'AllData' for A1:F500 and set print range
Task ID: calc_nrv_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Named range 'AllData' exists referencing Sheet1!$A$1:$F$500
  Component 2 (0.5): Print area set to $A$1:$F$500 on Sheet1
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_033'


def persist_app_state(domain: str):
    """Attempt to save any unsaved LibreOffice state via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 must exist
    if 'Sheet1' not in wb.sheetnames:
        print(f"FAIL: Sheet1 not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet1']

    # Component 1: Named range 'AllData' exists and refers to Sheet1!$A$1:$F$500 (0.5 points)
    try:
        # Look up 'AllData' in defined names (case-insensitive search)
        alldata_names = [n for n in wb.defined_names.values() if n.name.lower() == 'alldata']
        if len(alldata_names) > 0:
            alldata_value = alldata_names[0].attr_text
            # Normalize: remove quotes around sheet name if present
            normalized = alldata_value.replace("'", "").replace('"', '').upper()
            if normalized == "SHEET1!$A$1:$F$500":
                print(f"PASS: Component 1 — Named range 'AllData' = {alldata_value} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 — 'AllData' found but value is '{alldata_value}', expected 'Sheet1!$A$1:$F$500'")
        else:
            print(f"FAIL: Component 1 — Named range 'AllData' not found. Defined names: {list(wb.defined_names)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Print area set to $A$1:$F$500 on Sheet1 (0.5 points)
    try:
        print_area = ws.print_area
        if print_area:
            # print_area can be a string like "'Sheet1'!$A$1:$F$500" or "$A$1:$F$500"
            normalized_pa = str(print_area).replace("'", "").replace('"', '').upper()
            # Accept with or without sheet name prefix
            if "$A$1:$F$500" in normalized_pa:
                print(f"PASS: Component 2 — Print area = {print_area} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — Print area is '{print_area}', expected '$A$1:$F$500'")
        else:
            print(f"FAIL: Component 2 — No print area set on Sheet1")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
