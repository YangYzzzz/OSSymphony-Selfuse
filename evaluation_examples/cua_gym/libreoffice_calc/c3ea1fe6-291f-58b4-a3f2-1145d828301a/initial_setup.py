"""
Initial Setup: Major international sporting events spreadsheet with missing Host City data.
Task ID: osworld_multi_apps_conference_city_012
Domain: libreoffice_calc

Creates SportingEvents.xlsx with columns: Event, Year, Sport, Host City
Host City column is intentionally left blank (agent must fill it in via web research).
Rows are NOT sorted by year (agent must sort them).
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_012'
OUTPUT = f'{WORKDIR}/SportingEvents.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SportingEvents"

    # Column headers
    headers = ["Event", "Year", "Sport", "Host City"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Data rows — Host City is intentionally blank (task requirement)
    # Rows are NOT in chronological order (task requires sorting by Year)
    data = [
        ["Wimbledon Championships", 2018, "Tennis", ""],
        ["NBA Finals", 2016, "Basketball", ""],
        ["Super Bowl LIV", 2020, "American Football", ""],
        ["Tour de France Start", 2019, "Cycling", ""],
        ["ICC Cricket World Cup Final", 2019, "Cricket", ""],
    ]

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
