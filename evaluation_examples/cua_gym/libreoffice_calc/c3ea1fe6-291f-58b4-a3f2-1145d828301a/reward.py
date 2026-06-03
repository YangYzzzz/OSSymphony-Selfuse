"""
Reward Script: Fill in missing Host City data and sort rows by year in SportingEvents.xlsx
Task ID: osworld_multi_apps_conference_city_012
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: All Host City cells are filled (non-null)           — 0.4 points
  Component 2: Host City values match expected cities              — 0.4 points
  Component 3: Rows sorted chronologically by Year (ascending)    — 0.2 points
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_012'
FILE_PATH = f'{WORKDIR}/SportingEvents.xlsx'

# Expected host city mappings: (Event, Year) -> Host City
# Based on the task context (ground truth from task-gen agent)
EXPECTED_HOST_CITIES = {
    ('NBA Finals', 2016): 'Cleveland',
    ('Wimbledon Championships', 2018): 'London',
    ('Tour de France Start', 2019): 'Brussels',
    ('ICC Cricket World Cup Final', 2019): 'London',
    ('Super Bowl LIV', 2020): 'Miami',
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active/first sheet
    try:
        ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify the spreadsheet has the expected structure
    # Header check (not scored, just validate we have the right file)
    try:
        header = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        expected_header = ['Event', 'Year', 'Sport', 'Host City']
        if header != expected_header:
            print(f"CRITICAL: Unexpected headers: {header}, expected {expected_header}")
            print("REWARD: 0.0")
            return 0.0
        print(f"PRECONDITION OK: Headers correct: {header}")
    except Exception as e:
        print(f"CRITICAL: Cannot read headers: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: check we have 5 data rows (rows 2–6)
    try:
        data_rows = ws.max_row - 1  # excluding header
        if data_rows != 5:
            print(f"CRITICAL: Expected 5 data rows, found {data_rows}")
            print("REWARD: 0.0")
            return 0.0
        print(f"PRECONDITION OK: 5 data rows present")
    except Exception as e:
        print(f"CRITICAL: Cannot count rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Read all data rows
    try:
        rows = []
        for r in range(2, ws.max_row + 1):
            event = ws.cell(row=r, column=1).value
            year = ws.cell(row=r, column=2).value
            sport = ws.cell(row=r, column=3).value
            host_city = ws.cell(row=r, column=4).value
            rows.append((event, year, sport, host_city))
        print(f"DATA: Rows read: {rows}")
    except Exception as e:
        print(f"CRITICAL: Cannot read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All Host City cells are filled in (non-null) (0.4 points)
    # This FAILS on initial_env (all Host City = None) and PASSES on golden_env
    try:
        all_filled = all(row[3] is not None and str(row[3]).strip() != '' for row in rows)
        if all_filled:
            print(f"PASS: Component 1 — All Host City cells are filled (0.4 pts)")
            total_score += 0.4
        else:
            blank_events = [row[0] for row in rows if row[3] is None or str(row[3]).strip() == '']
            print(f"FAIL: Component 1 — Host City still blank for: {blank_events}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Host City values match expected cities (0.4 points)
    # Each correct mapping earns 0.08 points (5 events x 0.08 = 0.4)
    # This FAILS on initial_env (all blank) and PASSES on golden_env
    try:
        points_per_city = 0.4 / len(EXPECTED_HOST_CITIES)
        city_score = 0.0
        for row in rows:
            event, year, sport, host_city = row
            key = (event, year)
            if key in EXPECTED_HOST_CITIES:
                expected_city = EXPECTED_HOST_CITIES[key]
                if host_city and str(host_city).strip().lower() == expected_city.lower():
                    print(f"PASS: Component 2 — {event} ({year}): Host City = '{host_city}' (expected '{expected_city}')")
                    city_score += points_per_city
                else:
                    print(f"FAIL: Component 2 — {event} ({year}): expected '{expected_city}', got '{host_city}'")
            else:
                print(f"WARN: Component 2 — Unexpected event key: {key}")
        if city_score > 0:
            total_score += round(city_score, 4)
        print(f"Component 2 subtotal: {round(city_score, 4)}/0.4")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Rows are sorted chronologically by Year (ascending) (0.2 points)
    # This FAILS on initial_env (rows are out of order) and PASSES on golden_env
    try:
        years = [row[1] for row in rows]
        is_sorted = all(years[i] <= years[i + 1] for i in range(len(years) - 1))
        if is_sorted:
            print(f"PASS: Component 3 — Rows sorted by Year ascending: {years} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Rows NOT sorted by Year. Current order: {years}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
