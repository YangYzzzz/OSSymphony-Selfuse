"""
Initial Setup: Build a simple Sheet2 summary table showing total hours logged per project
Task ID: osworld_calc_sheet2_summary_table_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_sheet2_summary_table_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Timesheet ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Date', 'Employee Name', 'Project Code', 'Hours Logged']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic timesheet data — 20 rows across 5 projects
    data = [
        ('2025-01-06', 'Sarah Chen',       'PROJ-101', 6.0),
        ('2025-01-06', 'Marcus Johnson',   'PROJ-102', 8.0),
        ('2025-01-07', 'Priya Patel',      'PROJ-103', 7.5),
        ('2025-01-07', 'Derek Williams',   'PROJ-101', 5.5),
        ('2025-01-08', 'Sarah Chen',       'PROJ-104', 4.0),
        ('2025-01-08', 'Marcus Johnson',   'PROJ-103', 6.5),
        ('2025-01-09', 'Priya Patel',      'PROJ-102', 8.0),
        ('2025-01-09', 'Lucia Ramirez',    'PROJ-105', 7.0),
        ('2025-01-10', 'Derek Williams',   'PROJ-104', 3.5),
        ('2025-01-10', 'Sarah Chen',       'PROJ-101', 8.0),
        ('2025-01-13', 'Marcus Johnson',   'PROJ-105', 6.0),
        ('2025-01-13', 'Priya Patel',      'PROJ-101', 5.0),
        ('2025-01-14', 'Lucia Ramirez',    'PROJ-102', 7.5),
        ('2025-01-14', 'Derek Williams',   'PROJ-103', 8.0),
        ('2025-01-15', 'Sarah Chen',       'PROJ-105', 4.5),
        ('2025-01-15', 'Marcus Johnson',   'PROJ-104', 6.0),
        ('2025-01-16', 'Priya Patel',      'PROJ-105', 5.5),
        ('2025-01-16', 'Lucia Ramirez',    'PROJ-103', 7.0),
        ('2025-01-17', 'Derek Williams',   'PROJ-102', 6.5),
        ('2025-01-17', 'Sarah Chen',       'PROJ-104', 8.0),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # --- Sheet 2: Empty (agent will build the summary table here) ---
    ws2 = wb.create_sheet('Sheet2')
    # Sheet2 intentionally left empty — the agent must create the summary table

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
