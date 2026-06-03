"""
Reward Script: Build a simple Sheet2 summary table showing total hours logged per project
Task ID: osworld_calc_sheet2_summary_table_009
Domain: libreoffice_calc
Scoring:
  Component 1: Sheet2 has a header row with Project Code and total-hours column header (0.2 pts)
  Component 2: Sheet2 lists all 5 unique project codes from Sheet1 in column A (0.3 pts)
  Component 3: Sheet2 column B contains SUMIF formulas referencing Sheet1 data (0.5 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_009'

# Expected project codes derived from Sheet1 data
EXPECTED_PROJECTS = {'PROJ-101', 'PROJ-102', 'PROJ-103', 'PROJ-104', 'PROJ-105'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    The task requires Sheet2 to be populated with a summary table that:
    - Has a header row (Project Code | total hours column)
    - Lists each unique project code from Sheet1
    - Uses SUMIF formulas to calculate total hours per project from Sheet1
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in the workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws2 = wb['Sheet2']

    # Precondition: Sheet2 must have content (not empty)
    if ws2.max_row < 2 or ws2.max_column < 2:
        print(f"FAIL: Sheet2 appears empty or incomplete (max_row={ws2.max_row}, max_col={ws2.max_column})")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 1: Sheet2 has a header row with Project Code column and total hours column header (0.2 points)
    # This tests that the summary table structure was created (fails on initial empty Sheet2)
    try:
        header_a = ws2.cell(row=1, column=1).value
        header_b = ws2.cell(row=1, column=2).value

        has_project_header = (
            header_a is not None and
            'project' in str(header_a).lower() and
            'code' in str(header_a).lower()
        )
        has_hours_header = (
            header_b is not None and
            ('hour' in str(header_b).lower() or 'total' in str(header_b).lower())
        )

        if has_project_header and has_hours_header:
            print(f"PASS: Component 1 — Sheet2 has proper header row: A1='{header_a}', B1='{header_b}' (0.2 pts)")
            total_score += 0.2
        elif has_project_header or has_hours_header:
            print(f"FAIL: Component 1 — Partial header: A1='{header_a}', B1='{header_b}' (expected 'Project Code' and hours header)")
        else:
            print(f"FAIL: Component 1 — Header row missing or incorrect: A1='{header_a}', B1='{header_b}'")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check header row: {e}")

    # Component 2: Sheet2 lists all 5 unique project codes in column A (0.3 points)
    # This tests that all projects are enumerated in the summary table (fails on initial empty Sheet2)
    try:
        found_projects = set()
        for row_idx in range(2, ws2.max_row + 1):
            cell_val = ws2.cell(row=row_idx, column=1).value
            if cell_val is not None:
                found_projects.add(str(cell_val).strip())

        missing_projects = EXPECTED_PROJECTS - found_projects
        extra_projects = found_projects - EXPECTED_PROJECTS

        if not missing_projects and not extra_projects:
            print(f"PASS: Component 2 — All 5 project codes present in Sheet2 column A: {sorted(found_projects)} (0.3 pts)")
            total_score += 0.3
        elif not missing_projects and extra_projects:
            # All expected are present but there are extra rows — still acceptable if all expected are there
            print(f"PASS: Component 2 — All 5 expected project codes present (extra rows: {extra_projects}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Missing projects: {missing_projects}, Found: {found_projects}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check project codes: {e}")

    # Component 3: Sheet2 column B has SUMIF formulas referencing Sheet1 data (0.5 points)
    # This tests that the hours aggregation is formula-based using SUMIF (fails on initial empty Sheet2)
    try:
        sumif_count = 0
        total_data_rows = 0
        has_sheet1_ref = False

        for row_idx in range(2, ws2.max_row + 1):
            b_cell = ws2.cell(row=row_idx, column=2)
            b_val = b_cell.value
            if b_val is not None:
                total_data_rows += 1
                val_str = str(b_val).upper().replace(' ', '')
                # Check for SUMIF formula that references Sheet1
                if val_str.startswith('=SUMIF') and 'SHEET1' in val_str:
                    sumif_count += 1

        if sumif_count > 0 and sumif_count == len(EXPECTED_PROJECTS):
            print(f"PASS: Component 3 — All {sumif_count} data rows have SUMIF formulas referencing Sheet1 (0.5 pts)")
            total_score += 0.5
        elif sumif_count > 0:
            # Partial: some rows have SUMIF formulas referencing Sheet1
            partial_pts = round(0.5 * (sumif_count / len(EXPECTED_PROJECTS)), 4)
            print(f"PARTIAL: Component 3 — {sumif_count}/{len(EXPECTED_PROJECTS)} SUMIF rows found ({partial_pts} pts)")
            if partial_pts > 0:
                total_score += partial_pts
        else:
            # Check if there are any formulas at all (not SUMIF but still formula-based)
            formula_count = 0
            for row_idx in range(2, ws2.max_row + 1):
                b_val = ws2.cell(row=row_idx, column=2).value
                if b_val is not None and str(b_val).startswith('='):
                    formula_count += 1
            if formula_count > 0:
                print(f"FAIL: Component 3 — Found {formula_count} formulas but none are SUMIF referencing Sheet1 (0.0 pts)")
            else:
                print(f"FAIL: Component 3 — No SUMIF formulas found in Sheet2 column B (total_data_rows={total_data_rows}) (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check SUMIF formulas: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
