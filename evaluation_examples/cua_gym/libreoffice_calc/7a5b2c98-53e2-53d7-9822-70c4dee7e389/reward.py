"""
Reward Script: Create a professional table style with header formatting, alternating row colors, and borders
Task ID: calc_fmt_full_table_styling_065
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.35): Header row (A1:D1) has blue background (#4472C4)
  Component 2 (0.25): Header row font is white (#FFFFFF) and bold
  Component 3 (0.20): Even data rows (2,4,...,20) have light blue fill (#DEEAF1); odd rows have no fill
  Component 4 (0.20): All cells A1:D20 have thin borders on all sides
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_full_table_styling_065'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the 'Sales Report' sheet exists (precondition gate)
    if 'Sales Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Sales Report' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales Report']

    # Component 1: Header row (A1:D1) background is blue (#4472C4) (0.35 points)
    # The task requires A1:D1 to have background color #4472C4 (stored as FF4472C4 in ARGB)
    try:
        header_cols = range(1, 5)  # columns A to D
        header_fill_correct = 0
        header_fill_expected = 'FF4472C4'
        for col in header_cols:
            cell = ws.cell(row=1, column=col)
            actual_color = cell.fill.fgColor.rgb
            if actual_color == header_fill_expected:
                header_fill_correct += 1
        if header_fill_correct == 4:
            print(f"PASS: Component 1 — All 4 header cells (A1:D1) have blue background FF4472C4 (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Only {header_fill_correct}/4 header cells have correct blue fill FF4472C4")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row font is white (#FFFFFF) and bold (0.25 points)
    # Both the white color and bold attribute are required; partial sub-score:
    # 0.15 for bold on all 4 cells, 0.10 for white font color on all 4 cells
    try:
        bold_correct = 0
        white_font_correct = 0
        for col in range(1, 5):
            cell = ws.cell(row=1, column=col)
            if cell.font.bold is True:
                bold_correct += 1
            # White font may be stored as 00FFFFFF or FFFFFFFF
            try:
                font_color = cell.font.color.rgb
                if font_color in ('00FFFFFF', 'FFFFFFFF', 'ffffff', 'FFFFFF'):
                    white_font_correct += 1
            except Exception:
                pass

        if bold_correct == 4:
            print(f"PASS: Component 2a — All 4 header cells are bold (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2a — Only {bold_correct}/4 header cells are bold")

        if white_font_correct == 4:
            print(f"PASS: Component 2b — All 4 header cells have white font (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2b — Only {white_font_correct}/4 header cells have white font color")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Alternating row fills — even rows have #DEEAF1, odd rows have no fill (0.20 points)
    # This component checks the full alternating pattern: even rows filled + odd rows unfilled
    # The pattern only exists AFTER the task is done (initially ALL rows have no fill, so even rows
    # would fail the even-fill check — the combined correct alternating pattern is task-introduced).
    # We award 0.20 only when ALL even rows have the correct fill.
    # Odd rows staying unfilled is a natural consequence/correctness check, not awarded separately.
    try:
        even_fill_expected = 'FFDEEAF1'
        even_correct = 0
        even_rows = [r for r in range(2, 21) if r % 2 == 0]   # 2,4,...,20 — 10 rows
        odd_correct = 0
        odd_rows = [r for r in range(2, 21) if r % 2 != 0]     # 3,5,...,19 — 9 rows

        for row in even_rows:
            cell = ws.cell(row=row, column=1)  # check column A as representative
            if cell.fill.fgColor.rgb == even_fill_expected and cell.fill.fill_type == 'solid':
                even_correct += 1

        for row in odd_rows:
            cell = ws.cell(row=row, column=1)
            # No fill: fill_type should be None or fgColor is 00000000 (default)
            if cell.fill.fill_type is None or cell.fill.fill_type == 'none':
                odd_correct += 1

        # Score the full alternating pattern together — even rows with correct fill AND odd rows without fill
        # This compound check fails on initial (even rows have no fill initially, so even_correct==0)
        full_pattern_correct = (even_correct == len(even_rows) and odd_correct == len(odd_rows))
        even_only_correct = (even_correct == len(even_rows) and not full_pattern_correct)

        if full_pattern_correct:
            print(f"PASS: Component 3 — Correct alternating fills: {len(even_rows)} even rows FFDEEAF1, {len(odd_rows)} odd rows no fill (0.20 pts)")
            total_score += 0.20
        elif even_only_correct:
            # Partial: even rows correct but some odd rows have unexpected fills
            odd_issues = len(odd_rows) - odd_correct
            total_score += 0.10
            print(f"PASS: Component 3 partial — Even rows all have FFDEEAF1 fill, but {odd_issues} odd rows have unexpected fill (0.10 pts)")
        else:
            print(f"FAIL: Component 3 — Only {even_correct}/{len(even_rows)} even rows have correct fill FFDEEAF1")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: All cells A1:D20 have thin borders on all sides (0.20 points)
    try:
        border_correct = 0
        border_total = 0
        for row in range(1, 21):
            for col in range(1, 5):
                border_total += 1
                cell = ws.cell(row=row, column=col)
                sides = [
                    cell.border.left.style,
                    cell.border.right.style,
                    cell.border.top.style,
                    cell.border.bottom.style
                ]
                if all(s == 'thin' for s in sides):
                    border_correct += 1

        if border_correct == border_total:
            print(f"PASS: Component 4 — All {border_total} cells (A1:D20) have thin borders on all sides (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Only {border_correct}/{border_total} cells have complete thin borders")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
