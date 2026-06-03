"""
Initial Setup: Professional table styling task — Sales Report sheet
Task ID: calc_fmt_full_table_styling_065
Domain: libreoffice_calc

Creates a spreadsheet with one sheet 'Sales Report' containing header row and 19
data rows. NO formatting is applied — no fills, no borders, default font.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_full_table_styling_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Sales Report'

    # Headers (row 1) — no formatting
    headers = ['Date', 'Customer', 'Amount', 'Region']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 19 data rows (rows 2-20) — realistic sales data, no formatting
    data = [
        ['2025-01-03', 'Meridian Solutions',    12450.00, 'North'],
        ['2025-01-07', 'Apex Technologies',      8920.50, 'South'],
        ['2025-01-10', 'Blue Horizon Corp',     15300.75, 'East'],
        ['2025-01-14', 'Cedar Valley Ltd',       6780.00, 'West'],
        ['2025-01-17', 'Pinnacle Ventures',     22100.25, 'North'],
        ['2025-01-21', 'Sunrise Industries',     9640.00, 'South'],
        ['2025-01-24', 'Global Reach Inc',      18750.50, 'East'],
        ['2025-01-28', 'Northstar Partners',     5320.75, 'West'],
        ['2025-02-03', 'Clearwater Systems',    14200.00, 'North'],
        ['2025-02-07', 'Vantage Point LLC',     11880.25, 'South'],
        ['2025-02-11', 'Redwood Enterprises',    7450.00, 'East'],
        ['2025-02-14', 'Summit Group',          19600.50, 'West'],
        ['2025-02-18', 'Pacific Dynamics',       8900.75, 'North'],
        ['2025-02-21', 'Ironclad Corp',         25300.00, 'South'],
        ['2025-02-25', 'Lakeside Holdings',     10750.25, 'East'],
        ['2025-03-01', 'Cascade Industries',     6200.00, 'West'],
        ['2025-03-05', 'Sterling Associates',   17850.50, 'North'],
        ['2025-03-10', 'Keystone Partners',     13100.75, 'South'],
        ['2025-03-14', 'Horizon Digital',        9480.00, 'East'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Sales Report')
    print(f'  Rows: 1 header + 19 data rows (rows 2-20)')
    print(f'  Columns: Date, Customer, Amount, Region')
    print(f'  Formatting: NONE (no fills, no borders, default font)')


create_initial()
