"""
Reward Script: Diet and Nutrition Tracker
Task ID: calc_wf_063
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP formulas in Daily Log for auto-calculating macros (0.25)
  Component 2: SUM formulas in daily total rows (0.15)
  Component 3: Data validation (dropdown) on food column (0.15)
  Component 4: % of Target formulas with conditional formatting (0.15)
  Component 5: Summary sheet formulas (weekly avg, % of target, status) (0.15)
  Component 6: Stacked bar chart on Summary sheet (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_063'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    if 'Food Database' not in wb.sheetnames or 'Daily Log' not in wb.sheetnames:
        print("FAIL: Missing required sheets 'Food Database' or 'Daily Log'")
        print("REWARD: 0.0")
        return 0.0

    ws_db = wb['Food Database']
    ws_log = wb['Daily Log']

    # Find Summary sheet (may have different name)
    summary_sheet = None
    for sn in wb.sheetnames:
        if 'summary' in sn.lower():
            summary_sheet = wb[sn]
            break

    # =========================================================================
    # Component 1: VLOOKUP formulas in Daily Log for auto-calculating macros
    # (0.25 points)
    # In initial_env, columns D-G in food entry rows are None.
    # In golden_env, they contain VLOOKUP formulas.
    # =========================================================================
    try:
        vlookup_count = 0
        vlookup_expected = 0
        # Check food entry rows across 7 days (rows 6-13, 18-25, 30-37, etc.)
        day_start_rows = [6, 18, 30, 42, 54, 66, 78]
        for start_row in day_start_rows:
            for r in range(start_row, start_row + 8):
                food_cell = ws_log.cell(row=r, column=2).value
                if food_cell is not None:
                    vlookup_expected += 4  # D, E, F, G columns
                    for c in range(4, 8):  # columns D-G
                        val = ws_log.cell(row=r, column=c).value
                        if val is not None and isinstance(val, str) and 'VLOOKUP' in val.upper():
                            vlookup_count += 1

        if vlookup_expected > 0 and vlookup_count >= vlookup_expected * 0.8:
            print(f"PASS: Component 1 — VLOOKUP formulas found: {vlookup_count}/{vlookup_expected} (0.25 pts)")
            total_score += 0.25
        elif vlookup_count > 0:
            # Partial: at least some VLOOKUPs
            partial = 0.25 * (vlookup_count / max(vlookup_expected, 1))
            print(f"PARTIAL: Component 1 — VLOOKUP formulas: {vlookup_count}/{vlookup_expected} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas found in Daily Log food rows (D-G)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: SUM formulas in daily total rows (0.15 points)
    # In initial_env, total rows (14, 26, 38, 50, 62, 74, 86) columns D-G are None.
    # In golden_env, they contain SUM formulas.
    # =========================================================================
    try:
        total_rows = [14, 26, 38, 50, 62, 74, 86]
        sum_count = 0
        sum_expected = 7 * 4  # 7 days * 4 columns (D, E, F, G)
        for r in total_rows:
            for c in range(4, 8):
                val = ws_log.cell(row=r, column=c).value
                if val is not None and isinstance(val, str) and 'SUM' in val.upper():
                    sum_count += 1

        if sum_count >= sum_expected * 0.8:
            print(f"PASS: Component 2 — SUM formulas in daily totals: {sum_count}/{sum_expected} (0.15 pts)")
            total_score += 0.15
        elif sum_count > 0:
            partial = 0.15 * (sum_count / sum_expected)
            print(f"PARTIAL: Component 2 — SUM formulas: {sum_count}/{sum_expected} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No SUM formulas in daily total rows")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Data validation (dropdown) on food column (0.15 points)
    # In initial_env, no data validations exist.
    # In golden_env, list validation referencing Food Database on column B.
    # =========================================================================
    try:
        dv_list = ws_log.data_validations.dataValidation if ws_log.data_validations else []
        found_food_dv = False
        for dv in dv_list:
            if dv.type == 'list' and dv.formula1 is not None:
                formula_upper = str(dv.formula1).upper()
                if 'FOOD' in formula_upper or 'DATABASE' in formula_upper:
                    found_food_dv = True
                    break
                # Also check if it references column A of any sheet
                if '$A$' in formula_upper:
                    found_food_dv = True
                    break

        if found_food_dv:
            print(f"PASS: Component 3 — Food dropdown validation found referencing Food Database (0.15 pts)")
            total_score += 0.15
        elif len(dv_list) > 0:
            # Some data validation exists but may not reference Food Database
            print(f"PARTIAL: Component 3 — Data validation exists but may not reference Food Database (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — No data validation found on Daily Log")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: % of Target formulas and conditional formatting (0.15 points)
    # In initial_env, % of Target rows (15, 27, 39...) have None in D-G.
    # In golden_env, they have formulas and conditional formatting rules.
    # Split: 0.08 for formulas, 0.07 for conditional formatting.
    # =========================================================================
    try:
        pct_rows = [15, 27, 39, 51, 63, 75, 87]
        pct_formula_count = 0
        pct_expected = 7 * 4  # 7 days * 4 macro columns
        for r in pct_rows:
            for c in range(4, 8):
                val = ws_log.cell(row=r, column=c).value
                if val is not None and isinstance(val, str) and ('/' in val or '%' in val.upper() or '*100' in val):
                    pct_formula_count += 1

        if pct_formula_count >= pct_expected * 0.8:
            print(f"PASS: Component 4a — % of Target formulas: {pct_formula_count}/{pct_expected} (0.08 pts)")
            total_score += 0.08
        elif pct_formula_count > 0:
            partial = 0.08 * (pct_formula_count / pct_expected)
            print(f"PARTIAL: Component 4a — % of Target formulas: {pct_formula_count}/{pct_expected} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4a — No % of Target formulas found")

        # Check conditional formatting on Daily Log
        cf_count = len(list(ws_log.conditional_formatting))
        if cf_count >= 5:
            print(f"PASS: Component 4b — Conditional formatting rules: {cf_count} (0.07 pts)")
            total_score += 0.07
        elif cf_count > 0:
            partial = 0.07 * min(cf_count / 5.0, 1.0)
            print(f"PARTIAL: Component 4b — Conditional formatting rules: {cf_count} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4b — No conditional formatting on Daily Log")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Summary sheet formulas (0.15 points)
    # In initial_env, Summary B4-B7, D4-D7, E4-E7 are None.
    # In golden_env, they have AVERAGE, division, and IF formulas.
    # =========================================================================
    try:
        if summary_sheet is None:
            print(f"FAIL: Component 5 — No Summary sheet found")
        else:
            ws_sum = summary_sheet
            formula_count = 0
            # Check B4-B7 for AVERAGE formulas
            for r in range(4, 8):
                val = ws_sum.cell(row=r, column=2).value
                if val is not None and isinstance(val, str) and 'AVERAGE' in val.upper():
                    formula_count += 1

            # Check D4-D7 for % formulas
            for r in range(4, 8):
                val = ws_sum.cell(row=r, column=4).value
                if val is not None and isinstance(val, str) and ('/' in val or '*' in val):
                    formula_count += 1

            # Check E4-E7 for IF/status formulas
            for r in range(4, 8):
                val = ws_sum.cell(row=r, column=5).value
                if val is not None and isinstance(val, str) and 'IF' in val.upper():
                    formula_count += 1

            expected_formulas = 12  # 4 AVERAGE + 4 % + 4 IF
            if formula_count >= expected_formulas * 0.8:
                print(f"PASS: Component 5 — Summary formulas: {formula_count}/{expected_formulas} (0.15 pts)")
                total_score += 0.15
            elif formula_count > 0:
                partial = 0.15 * (formula_count / expected_formulas)
                print(f"PARTIAL: Component 5 — Summary formulas: {formula_count}/{expected_formulas} ({partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 5 — No formulas found in Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Stacked bar chart on Summary sheet (0.15 points)
    # In initial_env, no charts exist. In golden_env, stacked bar chart with 3 series.
    # Split: 0.08 for chart existence, 0.07 for stacked + correct series count.
    # =========================================================================
    try:
        if summary_sheet is None:
            print(f"FAIL: Component 6 — No Summary sheet found")
        else:
            ws_sum = summary_sheet
            charts = ws_sum._charts
            if len(charts) == 0:
                # Also check other sheets for charts
                all_charts = []
                for sn in wb.sheetnames:
                    all_charts.extend(wb[sn]._charts)
                if len(all_charts) > 0:
                    charts = all_charts
                    print(f"NOTE: Chart found on different sheet")

            if len(charts) >= 1:
                chart = charts[0]
                chart_found = True
                print(f"PASS: Component 6a — Chart exists on Summary sheet (0.08 pts)")
                total_score += 0.08

                # Check stacked grouping and series count
                is_stacked = False
                if hasattr(chart, 'grouping') and chart.grouping == 'stacked':
                    is_stacked = True
                elif hasattr(chart, 'grouping') and chart.grouping == 'percentStacked':
                    is_stacked = True

                has_multiple_series = len(chart.series) >= 3
                if is_stacked and has_multiple_series:
                    print(f"PASS: Component 6b — Stacked chart with {len(chart.series)} series (0.07 pts)")
                    total_score += 0.07
                elif is_stacked or has_multiple_series:
                    print(f"PARTIAL: Component 6b — stacked={is_stacked}, series={len(chart.series)} (0.035 pts)")
                    total_score += 0.035
                else:
                    print(f"FAIL: Component 6b — Not stacked ({getattr(chart, 'grouping', 'N/A')}) or insufficient series ({len(chart.series)})")
            else:
                print(f"FAIL: Component 6 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
