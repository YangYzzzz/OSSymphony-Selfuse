"""
Initial Setup: Diet and Nutrition Tracker
Task ID: calc_wf_063
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # ===== STYLING =====
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # ===== SHEET 1: Food Database =====
    ws_db = wb.active
    ws_db.title = "Food Database"

    db_headers = ["Food Name", "Serving Size", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)"]
    for c, h in enumerate(db_headers, 1):
        cell = ws_db.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 40 realistic foods
    foods = [
        ["Grilled Chicken Breast", "4 oz", 187, 35.0, 0.0, 4.1],
        ["Brown Rice", "1 cup cooked", 216, 5.0, 45.0, 1.8],
        ["Salmon Fillet", "4 oz", 233, 25.0, 0.0, 14.0],
        ["Broccoli", "1 cup", 55, 3.7, 11.2, 0.6],
        ["Sweet Potato", "1 medium", 103, 2.3, 24.0, 0.1],
        ["Greek Yogurt", "6 oz", 100, 17.0, 6.0, 0.7],
        ["Almonds", "1 oz (23 nuts)", 164, 6.0, 6.1, 14.2],
        ["Banana", "1 medium", 105, 1.3, 27.0, 0.4],
        ["Eggs (whole)", "2 large", 143, 12.6, 0.7, 9.5],
        ["Oatmeal", "1 cup cooked", 154, 5.3, 27.4, 2.6],
        ["Avocado", "1/2 medium", 161, 2.0, 8.6, 14.7],
        ["Turkey Breast", "4 oz", 153, 34.0, 0.0, 0.8],
        ["Quinoa", "1 cup cooked", 222, 8.1, 39.4, 3.6],
        ["Spinach (raw)", "2 cups", 14, 1.7, 2.2, 0.2],
        ["Cottage Cheese (2%)", "1 cup", 183, 24.0, 9.5, 5.0],
        ["Blueberries", "1 cup", 84, 1.1, 21.4, 0.5],
        ["Whole Wheat Bread", "2 slices", 138, 7.2, 23.6, 2.4],
        ["Peanut Butter", "2 tbsp", 188, 8.0, 6.0, 16.0],
        ["Tuna (canned in water)", "3 oz", 73, 16.5, 0.0, 0.6],
        ["Apple", "1 medium", 95, 0.5, 25.1, 0.3],
        ["Black Beans", "1 cup cooked", 227, 15.2, 40.8, 0.9],
        ["Chicken Thigh (skinless)", "4 oz", 209, 26.0, 0.0, 10.9],
        ["White Rice", "1 cup cooked", 206, 4.3, 44.5, 0.4],
        ["Mixed Greens Salad", "2 cups", 20, 1.6, 3.6, 0.2],
        ["Cheddar Cheese", "1 oz", 113, 7.0, 0.4, 9.3],
        ["Orange", "1 medium", 62, 1.2, 15.4, 0.2],
        ["Lentils", "1 cup cooked", 230, 17.9, 39.9, 0.8],
        ["Beef Sirloin", "4 oz", 207, 33.0, 0.0, 7.4],
        ["Pasta (whole wheat)", "1 cup cooked", 174, 7.5, 37.2, 0.8],
        ["Milk (2%)", "1 cup", 122, 8.1, 11.7, 4.8],
        ["Strawberries", "1 cup", 49, 1.0, 11.7, 0.5],
        ["Hummus", "2 tbsp", 70, 2.4, 6.0, 4.4],
        ["Shrimp", "4 oz", 120, 23.0, 1.0, 1.7],
        ["Granola", "1/2 cup", 210, 5.0, 34.0, 7.0],
        ["Celery with PB", "2 stalks + 1 tbsp PB", 110, 4.0, 5.0, 8.5],
        ["Protein Bar", "1 bar", 220, 20.0, 25.0, 8.0],
        ["Tofu (firm)", "4 oz", 88, 10.0, 2.2, 5.3],
        ["Corn Tortilla", "2 tortillas", 104, 2.8, 21.8, 1.4],
        ["Edamame", "1 cup shelled", 188, 18.5, 13.8, 8.1],
        ["Trail Mix", "1/4 cup", 175, 5.0, 15.0, 12.0],
    ]

    for r, row_data in enumerate(foods, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_db.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 3:
                cell.number_format = '0.0' if isinstance(val, float) else '0'
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws_db.column_dimensions["A"].width = 28
    ws_db.column_dimensions["B"].width = 20
    ws_db.column_dimensions["C"].width = 12
    ws_db.column_dimensions["D"].width = 14
    ws_db.column_dimensions["E"].width = 12
    ws_db.column_dimensions["F"].width = 12

    ws_db.freeze_panes = "A2"

    # ===== SHEET 2: Daily Log =====
    ws_log = wb.create_sheet("Daily Log")

    # --- Targets Section ---
    target_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    target_font = Font(name="Calibri", size=11, bold=True)

    ws_log.cell(row=1, column=1, value="Daily Nutrition Targets").font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    ws_log.merge_cells("A1:G1")
    ws_log["A1"].alignment = Alignment(horizontal="center")

    target_labels = ["", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)"]
    target_values = ["Targets:", 2000, 150, 250, 65]
    for c, (label, val) in enumerate(zip(target_labels, target_values), 1):
        lbl_cell = ws_log.cell(row=2, column=c, value=label if c == 1 else label)
        lbl_cell.font = header_font
        lbl_cell.fill = header_fill
        lbl_cell.alignment = header_align
        lbl_cell.border = thin_border

        val_cell = ws_log.cell(row=3, column=c, value=val)
        val_cell.font = target_font
        val_cell.fill = target_fill
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.border = thin_border
        if c >= 2:
            val_cell.number_format = '0'

    # --- Daily Log Entries ---
    log_headers = ["Date", "Food", "Servings", "Calories", "Protein (g)", "Carbs (g)", "Fat (g)"]
    days = ["Monday 2025-03-24", "Tuesday 2025-03-25", "Wednesday 2025-03-26",
            "Thursday 2025-03-27", "Friday 2025-03-28", "Saturday 2025-03-29",
            "Sunday 2025-03-30"]

    # Sample food entries for each day (indices into foods list)
    daily_meals = [
        # Monday
        [("Oatmeal", 1.5), ("Banana", 1), ("Grilled Chicken Breast", 1.5), ("Brown Rice", 1),
         ("Broccoli", 2), ("Greek Yogurt", 1), ("Almonds", 1), ("Apple", 1)],
        # Tuesday
        [("Eggs (whole)", 1), ("Whole Wheat Bread", 1), ("Turkey Breast", 1), ("Quinoa", 1),
         ("Spinach (raw)", 2), ("Salmon Fillet", 1), ("Sweet Potato", 1.5), ("Blueberries", 1)],
        # Wednesday
        [("Greek Yogurt", 1.5), ("Granola", 1), ("Black Beans", 1), ("Brown Rice", 1),
         ("Avocado", 1), ("Chicken Thigh (skinless)", 1), ("Mixed Greens Salad", 1), ("Orange", 1)],
        # Thursday
        [("Oatmeal", 1), ("Peanut Butter", 1), ("Tuna (canned in water)", 2), ("Whole Wheat Bread", 1),
         ("Cottage Cheese (2%)", 1), ("Beef Sirloin", 1), ("Sweet Potato", 1), ("Strawberries", 1.5)],
        # Friday
        [("Eggs (whole)", 1.5), ("Avocado", 0.5), ("Shrimp", 1.5), ("Pasta (whole wheat)", 1),
         ("Broccoli", 1.5), ("Tofu (firm)", 1.5), ("Edamame", 1), ("Trail Mix", 1)],
        # Saturday
        [("Protein Bar", 1), ("Banana", 1), ("Grilled Chicken Breast", 2), ("White Rice", 1),
         ("Hummus", 2), ("Lentils", 1), ("Corn Tortilla", 1), ("Milk (2%)", 1)],
        # Sunday
        [("Salmon Fillet", 1), ("Brown Rice", 1.5), ("Eggs (whole)", 1), ("Celery with PB", 1),
         ("Greek Yogurt", 1), ("Almonds", 1.5), ("Apple", 1), ("Chicken Thigh (skinless)", 1)],
    ]

    current_row = 5  # Start data after targets section

    day_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")

    for day_idx, (day_name, meals) in enumerate(zip(days, daily_meals)):
        # Day header row
        for c, h in enumerate(log_headers, 1):
            cell = ws_log.cell(row=current_row, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        current_row += 1

        # Food entries for this day
        for entry_idx, (food_name, servings) in enumerate(meals):
            ws_log.cell(row=current_row, column=1, value=day_name if entry_idx == 0 else "").border = thin_border
            ws_log.cell(row=current_row, column=2, value=food_name).border = thin_border
            cell_srv = ws_log.cell(row=current_row, column=3, value=servings)
            cell_srv.border = thin_border
            cell_srv.number_format = '0.0'
            cell_srv.alignment = Alignment(horizontal="center")

            # Columns D-G (Calories, Protein, Carbs, Fat) left EMPTY - task is to add VLOOKUP formulas
            for col in range(4, 8):
                ws_log.cell(row=current_row, column=col).border = thin_border

            current_row += 1

        # Daily totals row placeholder (empty - task is to add SUM formulas)
        totals_cell = ws_log.cell(row=current_row, column=1, value="Daily Total")
        totals_cell.font = Font(bold=True)
        totals_cell.fill = day_fill
        totals_cell.border = thin_border
        for col in range(2, 8):
            c = ws_log.cell(row=current_row, column=col)
            c.fill = day_fill
            c.border = thin_border

        current_row += 1

        # % of Target row placeholder (empty - task is to calculate)
        pct_cell = ws_log.cell(row=current_row, column=1, value="% of Target")
        pct_cell.font = Font(bold=True, italic=True)
        pct_cell.fill = day_fill
        pct_cell.border = thin_border
        for col in range(2, 8):
            c = ws_log.cell(row=current_row, column=col)
            c.fill = day_fill
            c.border = thin_border

        current_row += 1
        current_row += 1  # blank row between days

    # Column widths for Daily Log
    ws_log.column_dimensions["A"].width = 24
    ws_log.column_dimensions["B"].width = 28
    ws_log.column_dimensions["C"].width = 12
    ws_log.column_dimensions["D"].width = 12
    ws_log.column_dimensions["E"].width = 14
    ws_log.column_dimensions["F"].width = 12
    ws_log.column_dimensions["G"].width = 12

    # ===== SHEET 3: Summary =====
    ws_sum = wb.create_sheet("Summary")

    ws_sum.cell(row=1, column=1, value="Weekly Nutrition Summary").font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    ws_sum.merge_cells("A1:E1")
    ws_sum["A1"].alignment = Alignment(horizontal="center")

    sum_headers = ["Metric", "Weekly Avg", "Target", "% of Target", "Status"]
    for c, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=3, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    metrics = ["Calories", "Protein (g)", "Carbs (g)", "Fat (g)"]
    targets = [2000, 150, 250, 65]
    for r, (metric, target) in enumerate(zip(metrics, targets), 4):
        ws_sum.cell(row=r, column=1, value=metric).border = thin_border
        ws_sum.cell(row=r, column=2).border = thin_border  # Empty - to be calculated
        ws_sum.cell(row=r, column=3, value=target).border = thin_border
        ws_sum.cell(row=r, column=4).border = thin_border  # Empty - to be calculated
        ws_sum.cell(row=r, column=5).border = thin_border  # Empty - status

    ws_sum.column_dimensions["A"].width = 16
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 12
    ws_sum.column_dimensions["D"].width = 14
    ws_sum.column_dimensions["E"].width = 12

    # NO chart in initial - task requires creating chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
