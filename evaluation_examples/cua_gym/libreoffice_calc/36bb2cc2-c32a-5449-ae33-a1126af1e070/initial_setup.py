"""
Initial Setup: Quarterly Sales Report from Transaction Data
Task ID: calc_gen_report_039
Domain: libreoffice_calc

Creates a workbook with:
- 'Transactions' sheet: 1500 rows of realistic transaction data for Q1-Q4
- 'QuarterlyReport' sheet: exists but is completely empty
"""

import os
import random
import datetime
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_report_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Seed for reproducibility
random.seed(42)

# Realistic data pools
PRODUCT_LINES = ['Software', 'Hardware', 'Services', 'Support']

REPS = [
    'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
    'Jennifer Walsh', 'Robert Patel', 'Lisa Thompson', 'Kevin O\'Brien',
    'Amanda Foster', 'James Nakamura', 'Rachel Goldstein', 'Tyler Brooks',
    'Sophia Martinez', 'Nathan Clarke', 'Priya Sharma', 'Derek Hansen'
]

REGIONS = ['North', 'South', 'East', 'West', 'Central']

# Revenue ranges per product line (min, max per unit, units range)
PRODUCT_CONFIG = {
    'Software': {'price_range': (800, 5000), 'units_range': (1, 50)},
    'Hardware': {'price_range': (200, 2500), 'units_range': (1, 100)},
    'Services': {'price_range': (1500, 8000), 'units_range': (1, 20)},
    'Support': {'price_range': (300, 1200), 'units_range': (1, 30)},
}


def generate_transactions(n=1500):
    """Generate n realistic transaction rows."""
    rows = []
    # Generate dates spread evenly across Q1-Q4 of 2024
    start_date = datetime.date(2024, 1, 1)
    end_date = datetime.date(2024, 12, 31)
    date_range = (end_date - start_date).days

    for _ in range(n):
        # Random date in 2024
        offset = random.randint(0, date_range)
        tx_date = start_date + datetime.timedelta(days=offset)

        product_line = random.choice(PRODUCT_LINES)
        rep = random.choice(REPS)
        region = random.choice(REGIONS)

        cfg = PRODUCT_CONFIG[product_line]
        units = random.randint(*cfg['units_range'])
        price_per_unit = round(random.uniform(*cfg['price_range']), 2)
        revenue = round(units * price_per_unit, 2)

        rows.append([tx_date, product_line, rep, region, units, revenue])

    # Sort by date for natural ordering
    rows.sort(key=lambda x: x[0])
    return rows


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Transactions ----
    ws_tx = wb.active
    ws_tx.title = 'Transactions'

    # Headers
    headers = ['Date', 'Product Line', 'Rep', 'Region', 'Units', 'Revenue']
    for col, h in enumerate(headers, 1):
        cell = ws_tx.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Generate and write 1500 transaction rows
    transactions = generate_transactions(1500)
    for r, row_data in enumerate(transactions, 2):
        for c, val in enumerate(row_data, 1):
            ws_tx.cell(row=r, column=c, value=val)

    # Format the Date column
    for row in range(2, 1502):
        ws_tx.cell(row=row, column=1).number_format = 'yyyy-mm-dd'

    # Format Revenue column
    for row in range(2, 1502):
        ws_tx.cell(row=row, column=6).number_format = '$#,##0.00'

    # Column widths
    ws_tx.column_dimensions['A'].width = 14
    ws_tx.column_dimensions['B'].width = 16
    ws_tx.column_dimensions['C'].width = 22
    ws_tx.column_dimensions['D'].width = 12
    ws_tx.column_dimensions['E'].width = 10
    ws_tx.column_dimensions['F'].width = 16

    # ---- Sheet 2: QuarterlyReport (empty) ----
    ws_qr = wb.create_sheet('QuarterlyReport')
    # This sheet is intentionally left empty — the agent must build it

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Transactions sheet: 1500 rows of transaction data')
    print(f'  QuarterlyReport sheet: empty (agent must populate)')


create_initial()
