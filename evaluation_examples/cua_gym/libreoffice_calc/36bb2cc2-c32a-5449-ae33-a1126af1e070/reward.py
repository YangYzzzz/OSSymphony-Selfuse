"""
Reward Script: Quarterly Sales Report Generation
Task ID: calc_gen_report_039
Domain: libreoffice_calc
Scoring:
  - Component 1: Monthly revenue table structure (0.30)
    - 12-month rows (Jan-Dec in rows 3-14), column headers present
    - Grand total column (F) and annual total row (15)
  - Component 2: SUMIFS formulas for product line data (0.25)
    - B3:E14 have SUMIFS formulas referencing Transactions sheet
    - F3:F14 have SUM formulas (grand total column)
  - Component 3: QoQ growth rows (0.15)
    - Rows 16-18 contain QoQ growth formulas
    - Uses correct quarter references (Q1=rows3-5, Q2=rows6-8, Q3=rows9-11, Q4=rows12-14)
  - Component 4: Header row formatting (0.15)
    - Row 2 has dark navy fill and bold white text
  - Component 5: Alternating row colors + currency formatting (0.10)
    - Odd data rows (3,5,7,...) white, even data rows (4,6,8,...) light gray
    - Revenue cells have currency format ($#,##0.00)
  - Component 6: Chart present in QuarterlyReport (0.05)
    - At least 1 chart with at least 1 series
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_report_039'

MONTHS = ['January', 'February', 'March', 'April', 'May', 'June',
          'July', 'August', 'September', 'October', 'November', 'December']
PRODUCT_LINES = ['Software', 'Hardware', 'Services', 'Support']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — gate check
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: QuarterlyReport sheet must exist
    if 'QuarterlyReport' not in wb.sheetnames:
        print("FAIL: 'QuarterlyReport' sheet not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['QuarterlyReport']

    # Component 1: Monthly revenue table structure (0.30 points)
    # Checks: 12 month rows (A3:A14), column headers in row 2,
    # Annual Total in row 15, Total column header in F2
    try:
        # Check column headers in row 2
        expected_headers = ['Month', 'Software', 'Hardware', 'Services', 'Support', 'Total']
        headers_found = []
        for col, expected in enumerate(expected_headers, 1):
            val = ws.cell(row=2, column=col).value
            headers_found.append(val == expected)

        correct_headers = sum(headers_found)

        # Check 12 month names in A3:A14
        months_found = 0
        for i, month in enumerate(MONTHS):
            val = ws.cell(row=3 + i, column=1).value
            if val == month:
                months_found += 1

        # Check Annual Total row label in A15
        annual_total_label = ws.cell(row=15, column=1).value
        has_annual_total = (annual_total_label is not None and
                            'total' in str(annual_total_label).lower())

        # Award points if structure is substantially complete
        structure_score = 0.0
        if correct_headers >= 5:  # at least 5 of 6 headers correct
            structure_score += 0.10
            print(f"PASS: Column headers present ({correct_headers}/6 correct)")
        else:
            print(f"FAIL: Column headers incomplete ({correct_headers}/6 correct, found: {[ws.cell(row=2, column=c+1).value for c in range(6)]})")

        if months_found >= 10:  # at least 10 of 12 months present
            structure_score += 0.15
            print(f"PASS: Month rows present ({months_found}/12 months found in A3:A14)")
        else:
            print(f"FAIL: Month rows incomplete ({months_found}/12 found in A3:A14)")

        if has_annual_total:
            structure_score += 0.05
            print(f"PASS: Annual total row label present in A15 ('{annual_total_label}')")
        else:
            print(f"FAIL: Annual total row label missing in A15 (found: {repr(ws.cell(row=15, column=1).value)})")

        total_score += structure_score
        print(f"  Component 1 subtotal: {structure_score}/0.30")

    except Exception as e:
        print(f"ERROR: Component 1 (table structure) — {e}")

    # Component 2: SUMIFS formulas for product line data (0.25 points)
    # Checks: B3:E14 use SUMIFS referencing Transactions sheet,
    # F3:F14 use SUM, B15:E15 use SUM for annual totals
    try:
        sumifs_count = 0
        total_cells = 0  # B3:E14 = 4 cols x 12 rows = 48 cells
        for row in range(3, 15):
            for col in range(2, 6):  # columns B-E
                val = ws.cell(row=row, column=col).value
                total_cells += 1
                if isinstance(val, str) and 'SUMIFS' in val.upper() and 'Transactions' in val:
                    sumifs_count += 1

        # Check SUM formulas in F3:F14 (grand total column)
        sum_total_count = 0
        for row in range(3, 15):
            val = ws.cell(row=row, column=6).value
            if isinstance(val, str) and val.upper().startswith('=SUM'):
                sum_total_count += 1

        # Check annual total row formulas B15:E15
        annual_sum_count = 0
        for col in range(2, 7):  # B15:F15
            val = ws.cell(row=15, column=col).value
            if isinstance(val, str) and 'SUM' in val.upper():
                annual_sum_count += 1

        formula_score = 0.0
        if sumifs_count >= 40:  # at least 40 of 48 SUMIFS cells
            formula_score += 0.15
            print(f"PASS: SUMIFS formulas in product line cells ({sumifs_count}/48 cells)")
        elif sumifs_count >= 20:
            formula_score += 0.07
            print(f"PARTIAL: Some SUMIFS formulas present ({sumifs_count}/48 cells)")
        else:
            print(f"FAIL: Insufficient SUMIFS formulas ({sumifs_count}/48 cells)")

        if sum_total_count >= 10:
            formula_score += 0.05
            print(f"PASS: SUM formulas in grand total column F ({sum_total_count}/12)")
        else:
            print(f"FAIL: Missing SUM formulas in column F ({sum_total_count}/12)")

        if annual_sum_count >= 4:
            formula_score += 0.05
            print(f"PASS: Annual total SUM formulas in row 15 ({annual_sum_count}/5)")
        else:
            print(f"FAIL: Missing annual total SUM formulas in row 15 ({annual_sum_count}/5)")

        total_score += formula_score
        print(f"  Component 2 subtotal: {formula_score}/0.25")

    except Exception as e:
        print(f"ERROR: Component 2 (SUMIFS formulas) — {e}")

    # Component 3: QoQ growth rows (0.15 points)
    # Rows 16-18 should have QoQ growth formulas with quarter-based SUM references
    try:
        qoq_formula_count = 0
        qoq_label_count = 0

        for row in range(16, 19):
            # Check label in column A
            label_val = ws.cell(row=row, column=1).value
            if label_val is not None and 'qoq' in str(label_val).lower():
                qoq_label_count += 1

            # Check formulas in B-E (cols 2-5)
            for col in range(2, 6):
                val = ws.cell(row=row, column=col).value
                # QoQ formulas should reference SUM of 3-month ranges
                if isinstance(val, str) and 'SUM' in val.upper():
                    qoq_formula_count += 1

        qoq_score = 0.0
        if qoq_label_count >= 2:
            qoq_score += 0.05
            print(f"PASS: QoQ growth row labels present ({qoq_label_count}/3 rows labeled)")
        else:
            print(f"FAIL: QoQ growth row labels missing ({qoq_label_count}/3 rows with 'QoQ' label)")

        if qoq_formula_count >= 8:  # at least 8 of 12 formula cells present
            qoq_score += 0.10
            print(f"PASS: QoQ growth formulas present ({qoq_formula_count}/12 formula cells)")
        elif qoq_formula_count >= 4:
            qoq_score += 0.05
            print(f"PARTIAL: Some QoQ growth formulas ({qoq_formula_count}/12 formula cells)")
        else:
            print(f"FAIL: QoQ growth formulas missing ({qoq_formula_count}/12)")

        total_score += qoq_score
        print(f"  Component 3 subtotal: {qoq_score}/0.15")

    except Exception as e:
        print(f"ERROR: Component 3 (QoQ growth rows) — {e}")

    # Component 4: Header row formatting — navy fill + white bold text (0.15 points)
    try:
        # Check row 2 header formatting
        header_cells_with_navy = 0
        header_cells_with_bold = 0
        header_cells_with_white_font = 0
        cells_checked = 0

        for col in range(1, 7):  # A2:F2
            cell = ws.cell(row=2, column=col)
            cells_checked += 1
            try:
                fill_rgb = cell.fill.fgColor.rgb
                # Navy/dark blue shades: check if fill is dark (not white/light)
                # Accept any dark blue: FF1F3864, FF003366, FF002060, etc.
                # Strategy: the red+green+blue components are all relatively low
                if len(fill_rgb) == 8:
                    r = int(fill_rgb[2:4], 16)
                    g = int(fill_rgb[4:6], 16)
                    b = int(fill_rgb[6:8], 16)
                    # Dark color: brightness < 100 (out of 255)
                    brightness = (r + g + b) / 3
                    if brightness < 100 and b > r:  # dark and blue-dominant
                        header_cells_with_navy += 1
            except Exception:
                pass

            if cell.font.bold:
                header_cells_with_bold += 1

            try:
                font_rgb = cell.font.color.rgb
                if font_rgb and len(font_rgb) >= 6:
                    # White or near-white font: all components > 200
                    r = int(font_rgb[-6:-4], 16)
                    g = int(font_rgb[-4:-2], 16)
                    b = int(font_rgb[-2:], 16)
                    if r > 200 and g > 200 and b > 200:
                        header_cells_with_white_font += 1
            except Exception:
                pass

        format_score = 0.0
        if header_cells_with_navy >= 4:
            format_score += 0.07
            print(f"PASS: Header row has dark navy fill ({header_cells_with_navy}/6 cells)")
        else:
            print(f"FAIL: Header row navy fill missing ({header_cells_with_navy}/6 cells have dark fill)")

        if header_cells_with_bold >= 4:
            format_score += 0.04
            print(f"PASS: Header row has bold text ({header_cells_with_bold}/6 cells)")
        else:
            print(f"FAIL: Header row bold missing ({header_cells_with_bold}/6 bold cells)")

        if header_cells_with_white_font >= 4:
            format_score += 0.04
            print(f"PASS: Header row has white font ({header_cells_with_white_font}/6 cells)")
        else:
            print(f"FAIL: Header row white font missing ({header_cells_with_white_font}/6 cells)")

        total_score += format_score
        print(f"  Component 4 subtotal: {format_score}/0.15")

    except Exception as e:
        print(f"ERROR: Component 4 (header formatting) — {e}")

    # Component 5: Alternating row colors + currency formatting (0.10 points)
    try:
        # Odd rows (3,5,7,9,11,13) should be white (FFFFFFFF)
        # Even rows (4,6,8,10,12,14) should be light gray (FFF2F2F2 or similar)
        alternating_correct = 0
        total_data_rows = 12

        for i, row in enumerate(range(3, 15)):
            cell = ws.cell(row=row, column=2)  # use column B as representative
            try:
                fill_rgb = cell.fill.fgColor.rgb
                if len(fill_rgb) == 8:
                    r = int(fill_rgb[2:4], 16)
                    g = int(fill_rgb[4:6], 16)
                    b_val = int(fill_rgb[6:8], 16)
                    # White: all channels >= 252 (covers FF/255)
                    is_white = (r >= 252 and g >= 252 and b_val >= 252)
                    # Light gray: all channels between 200-251 and roughly equal
                    # (e.g. F2F2F2 = 242,242,242 — near white but not 255)
                    is_gray = (200 <= r <= 251 and 200 <= g <= 251 and 200 <= b_val <= 251
                               and abs(r - g) < 30 and abs(g - b_val) < 30)

                    if (i % 2 == 0 and is_white) or (i % 2 == 1 and is_gray):
                        alternating_correct += 1
            except Exception:
                pass

        # Check currency format in B3:F14 (revenue cells)
        currency_cells = 0
        currency_total = 0
        for row in range(3, 15):
            for col in range(2, 7):  # B-F
                currency_total += 1
                fmt = ws.cell(row=row, column=col).number_format
                if fmt and ('$' in fmt or '0.00' in fmt or '#,##0' in fmt):
                    currency_cells += 1

        alt_score = 0.0
        if alternating_correct >= 8:
            alt_score += 0.06
            print(f"PASS: Alternating row colors correct ({alternating_correct}/12 rows)")
        elif alternating_correct >= 4:
            alt_score += 0.03
            print(f"PARTIAL: Some alternating row colors ({alternating_correct}/12 rows)")
        else:
            print(f"FAIL: Alternating row colors not applied ({alternating_correct}/12 rows correct)")

        if currency_cells >= 40:  # at least 40 of 60 revenue cells have currency format
            alt_score += 0.04
            print(f"PASS: Currency formatting on revenue cells ({currency_cells}/{currency_total} cells)")
        elif currency_cells >= 20:
            alt_score += 0.02
            print(f"PARTIAL: Some currency formatting ({currency_cells}/{currency_total} cells)")
        else:
            print(f"FAIL: Currency formatting missing ({currency_cells}/{currency_total} cells)")

        total_score += alt_score
        print(f"  Component 5 subtotal: {alt_score}/0.10")

    except Exception as e:
        print(f"ERROR: Component 5 (alternating colors + currency) — {e}")

    # Component 6: Chart present in QuarterlyReport (0.05 points)
    try:
        charts = ws._charts
        if len(charts) >= 1 and len(charts[0].series) >= 1:
            total_score += 0.05
            print(f"PASS: Chart present in QuarterlyReport ({len(charts)} chart(s), "
                  f"{len(charts[0].series)} series)")
        elif len(charts) >= 1:
            total_score += 0.02
            print(f"PARTIAL: Chart present but no series ({len(charts)} chart(s))")
        else:
            print("FAIL: No chart found in QuarterlyReport sheet")

    except Exception as e:
        print(f"ERROR: Component 6 (chart) — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
