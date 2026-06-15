"""
Initial Setup: Personal Financial Loan Comparison Calculator
Task ID: calc_grs_020
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loan Comparison"

    # ---- Styling definitions ----
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    label_font = Font(name="Calibri", size=11, bold=True)
    section_font = Font(name="Calibri", size=12, bold=True, color="2F5496")
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # ---- Column widths ----
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # ---- Row 1: Title (merged) ----
    ws.merge_cells("A1:D1")
    ws["A1"] = "Auto Loan Comparison Calculator"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # ---- Row 2: Blank spacer ----

    # ---- Row 3: Input Section Header ----
    input_headers = ["", "Loan Offer 1", "Loan Offer 2", "Loan Offer 3"]
    for c, h in enumerate(input_headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        if c > 1:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        cell.border = thin_border

    # ---- Rows 4-8: Input Data ----
    # Row 4: Bank Name
    ws.cell(row=4, column=1, value="Bank Name").font = label_font
    ws.cell(row=4, column=2, value="First National Bank")
    ws.cell(row=4, column=3, value="City Credit Union")
    ws.cell(row=4, column=4, value="AutoFin Direct")

    # Row 5: Loan Amount
    ws.cell(row=5, column=1, value="Loan Amount ($)").font = label_font
    ws.cell(row=5, column=2, value=28000)
    ws.cell(row=5, column=3, value=28000)
    ws.cell(row=5, column=4, value=28000)
    for c in range(2, 5):
        ws.cell(row=5, column=c).number_format = '$#,##0.00'

    # Row 6: Annual Interest Rate
    ws.cell(row=6, column=1, value="Annual Interest Rate (%)").font = label_font
    ws.cell(row=6, column=2, value=0.059)   # 5.9%
    ws.cell(row=6, column=3, value=0.045)   # 4.5%
    ws.cell(row=6, column=4, value=0.068)   # 6.8%
    for c in range(2, 5):
        ws.cell(row=6, column=c).number_format = '0.00%'

    # Row 7: Loan Term (years)
    ws.cell(row=7, column=1, value="Loan Term (years)").font = label_font
    ws.cell(row=7, column=2, value=5)
    ws.cell(row=7, column=3, value=5)
    ws.cell(row=7, column=4, value=5)

    # Row 8: Origination Fee
    ws.cell(row=8, column=1, value="Origination Fee ($)").font = label_font
    ws.cell(row=8, column=2, value=0)
    ws.cell(row=8, column=3, value=250)
    ws.cell(row=8, column=4, value=500)
    for c in range(2, 5):
        ws.cell(row=8, column=c).number_format = '$#,##0.00'

    # Apply borders to input section
    for r in range(4, 9):
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = thin_border

    # ---- Row 9: Blank spacer ----

    # ---- Row 10: Calculated Results Section Header ----
    calc_headers = ["Calculated Results", "Loan Offer 1", "Loan Offer 2", "Loan Offer 3"]
    for c, h in enumerate(calc_headers, 1):
        cell = ws.cell(row=10, column=c, value=h)
        if c == 1:
            cell.font = section_font
        else:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        cell.border = thin_border

    # ---- Rows 11-14: Empty calculated fields (agent must fill these) ----
    calc_labels = [
        "Monthly Payment",
        "Total Amount Paid",
        "Total Interest Paid",
        "Effective APR (incl. fees)",
    ]
    for r, label in enumerate(calc_labels, 11):
        ws.cell(row=r, column=1, value=label).font = label_font
        ws.cell(row=r, column=1).border = thin_border
        for c in range(2, 5):
            ws.cell(row=r, column=c).border = thin_border

    # Set number formats for calculated fields (hint for agent)
    for c in range(2, 5):
        ws.cell(row=11, column=c).number_format = '$#,##0.00'
        ws.cell(row=12, column=c).number_format = '$#,##0.00'
        ws.cell(row=13, column=c).number_format = '$#,##0.00'
        ws.cell(row=14, column=c).number_format = '0.00%'

    # ---- Row 16: Instructions note ----
    ws.merge_cells("A16:D16")
    ws["A16"] = "Note: Fill in the Calculated Results section using formulas. Use PMT for monthly payments."
    ws["A16"].font = Font(name="Calibri", size=10, italic=True, color="808080")

    # Freeze panes to keep headers visible
    ws.freeze_panes = "A3"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
