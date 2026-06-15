"""
Reward Script: Personal Financial Loan Comparison Calculator
Task ID: calc_grs_020
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): PMT formulas in Monthly Payment row (B11:D11)
  Component 2 (0.25): Calculated fields formulas (Total Amount, Total Interest, Effective APR)
  Component 3 (0.15): Data validation on interest rate cells (B6:D6)
  Component 4 (0.15): Conditional formatting (green lowest payment, blue lowest interest)
  Component 5 (0.20): Bar chart with 2 series for comparison
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_020'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the main sheet (should be "Loan Comparison" or similar)
    ws = None
    for sn in wb.sheetnames:
        ws_candidate = wb[sn]
        # Look for sheet with loan data
        if ws_candidate['A4'] and ws_candidate['A4'].value and 'bank' in str(ws_candidate['A4'].value).lower():
            ws = ws_candidate
            break
    if ws is None:
        ws = wb.worksheets[0]

    # Component 1: PMT formulas in Monthly Payment row B11:D11 (0.25 points)
    # In initial_env these cells are empty. In golden they contain PMT formulas.
    try:
        pmt_count = 0
        for col in ['B', 'C', 'D']:
            cell_ref = f'{col}11'
            val = ws[cell_ref].value
            if val is not None and isinstance(val, str) and 'PMT' in val.upper():
                pmt_count += 1
                print(f"  FOUND: {cell_ref} has PMT formula: {val}")
            elif val is not None and isinstance(val, (int, float)):
                # May have been computed -- check if it's a reasonable monthly payment
                # (should be a positive number in range ~400-700 for $28k loans)
                if 100 < float(val) < 2000:
                    pmt_count += 1
                    print(f"  FOUND: {cell_ref} has computed value: {val}")
                else:
                    print(f"  FAIL: {cell_ref} has unexpected value: {val}")
            else:
                print(f"  FAIL: {cell_ref} is empty or unexpected: {val}")

        if pmt_count == 3:
            print(f"PASS: Component 1 - All 3 PMT formulas present (0.25 pts)")
            total_score += 0.25
        elif pmt_count > 0:
            partial = round(0.25 * pmt_count / 3, 2)
            print(f"PARTIAL: Component 1 - {pmt_count}/3 PMT formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No PMT formulas found in B11:D11")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Other calculated fields formulas in rows 12-14 (0.25 points)
    # Total Amount Paid (row 12), Total Interest Paid (row 13), Effective APR (row 14)
    # In initial_env these are all empty. In golden they contain formulas.
    try:
        calc_count = 0
        total_expected = 9  # 3 columns x 3 rows

        for row_num in [12, 13, 14]:
            for col in ['B', 'C', 'D']:
                cell_ref = f'{col}{row_num}'
                val = ws[cell_ref].value
                if val is not None:
                    if isinstance(val, str) and val.startswith('='):
                        calc_count += 1
                    elif isinstance(val, (int, float)):
                        # Computed value present
                        calc_count += 1
                    else:
                        print(f"  INFO: {cell_ref} has non-formula text: {val}")
                else:
                    print(f"  FAIL: {cell_ref} is empty")

        if calc_count >= 9:
            print(f"PASS: Component 2 - All 9 calculated fields present (0.25 pts)")
            total_score += 0.25
        elif calc_count > 0:
            partial = round(0.25 * calc_count / 9, 2)
            print(f"PARTIAL: Component 2 - {calc_count}/9 calculated fields ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No calculated fields found in rows 12-14")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Data validation on interest rate cells (0.15 points)
    # In initial_env there is no data validation. In golden, B6:D6 has decimal validation.
    try:
        dv_found = False
        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                sqref_str = str(dv.sqref)
                # Check if any of the interest rate cells (B6, C6, D6) are covered
                covers_rate_cells = False
                for col in ['B', 'C', 'D']:
                    if f'{col}6' in sqref_str:
                        covers_rate_cells = True
                        break
                # Also check range notation like B6:D6
                if 'B6:D6' in sqref_str or 'B6' in sqref_str:
                    covers_rate_cells = True

                if covers_rate_cells:
                    # Check type is decimal/whole/custom (any numeric validation)
                    if dv.type in ('decimal', 'whole', 'custom', 'list'):
                        dv_found = True
                        print(f"  FOUND: Data validation type={dv.type}, formula1={dv.formula1}, formula2={dv.formula2}, sqref={dv.sqref}")

        if dv_found:
            print(f"PASS: Component 3 - Data validation on interest rate cells (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 - No data validation found on interest rate cells (B6:D6)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Conditional formatting (0.15 points)
    # In initial_env there is no conditional formatting.
    # In golden: green on lowest monthly payment (row 11), blue on lowest total interest (row 13)
    try:
        cf_payment = False
        cf_interest = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                # Check if this CF covers row 11 (monthly payment)
                if '11' in cf_range:
                    cf_payment = True
                    print(f"  FOUND: CF on monthly payment range: {cf_range}")
                # Check if this CF covers row 13 (total interest)
                if '13' in cf_range:
                    cf_interest = True
                    print(f"  FOUND: CF on total interest range: {cf_range}")

        cf_count = sum([cf_payment, cf_interest])
        if cf_count == 2:
            print(f"PASS: Component 4 - Both conditional formatting rules present (0.15 pts)")
            total_score += 0.15
        elif cf_count == 1:
            print(f"PARTIAL: Component 4 - {cf_count}/2 conditional formatting rules (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 4 - No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Bar chart with comparison data (0.20 points)
    # In initial_env there are no charts. In golden there is a column chart with 2 series.
    try:
        charts = ws._charts
        if len(charts) == 0:
            # Check all sheets for charts
            for sn in wb.sheetnames:
                sheet = wb[sn]
                if len(sheet._charts) > 0:
                    charts = sheet._charts
                    break

        if len(charts) >= 1:
            chart = charts[0]
            chart_type = chart.type  # "col" or "bar"
            series_count = len(chart.series)

            # Check chart type is bar/column
            if chart_type in ('col', 'bar'):
                if series_count >= 2:
                    print(f"PASS: Component 5 - Bar chart with {series_count} series (0.20 pts)")
                    total_score += 0.20
                elif series_count == 1:
                    print(f"PARTIAL: Component 5 - Bar chart found but only {series_count} series (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"PARTIAL: Component 5 - Bar chart found but no series (0.05 pts)")
                    total_score += 0.05
            else:
                # Chart exists but wrong type -- partial credit
                if series_count >= 2:
                    print(f"PARTIAL: Component 5 - Chart type={chart_type} (not bar/col) with {series_count} series (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"PARTIAL: Component 5 - Chart type={chart_type} with {series_count} series (0.05 pts)")
                    total_score += 0.05
        else:
            print(f"FAIL: Component 5 - No charts found in workbook")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state('libreoffice_calc')

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
