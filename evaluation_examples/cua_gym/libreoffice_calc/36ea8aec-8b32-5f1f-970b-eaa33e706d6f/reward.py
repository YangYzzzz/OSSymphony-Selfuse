"""
Reward Script: Weighted Average Variance by Department using SUMPRODUCT
Task ID: calc_gg5_034
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.15): E1 header labeled 'Weighted Avg Variance'
  - Component 2 (0.35): E2:E8 all contain SUMPRODUCT formulas
  - Component 3 (0.30): Formulas reference Detail sheet columns correctly
  - Component 4 (0.20): Formula structure implements weighted average pattern
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_034'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Analysis sheet must exist
    if 'Analysis' not in wb.sheetnames:
        print("CRITICAL: 'Analysis' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Analysis']

    # Component 1: E1 header is labeled 'Weighted Avg Variance' (0.15 points)
    try:
        e1_val = ws['E1'].value
        if e1_val and 'weighted' in str(e1_val).lower() and 'variance' in str(e1_val).lower():
            print(f"PASS: Component 1 — E1 header is '{e1_val}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — E1 expected 'Weighted Avg Variance', found: {e1_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2:E8 all contain SUMPRODUCT formulas (0.35 points)
    # Each cell gets ~0.05 pts; all 7 must have SUMPRODUCT to get full points
    try:
        sumproduct_count = 0
        for row in range(2, 9):
            cell_val = ws.cell(row=row, column=5).value
            if cell_val and isinstance(cell_val, str) and 'SUMPRODUCT' in cell_val.upper():
                sumproduct_count += 1
            else:
                print(f"  INFO: E{row} value: {cell_val}")

        if sumproduct_count == 7:
            print(f"PASS: Component 2 — All 7 cells E2:E8 contain SUMPRODUCT formulas (0.35 pts)")
            total_score += 0.35
        elif sumproduct_count > 0:
            partial = round(0.35 * (sumproduct_count / 7), 2)
            print(f"PARTIAL: Component 2 — {sumproduct_count}/7 cells have SUMPRODUCT ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No SUMPRODUCT formulas found in E2:E8")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas reference Detail sheet columns correctly (0.30 points)
    # Expected: formulas reference Detail!A (Department), Detail!E (Variance), Detail!C (Budget Amount)
    try:
        detail_ref_count = 0
        for row in range(2, 9):
            cell_val = ws.cell(row=row, column=5).value
            if cell_val and isinstance(cell_val, str):
                formula_upper = cell_val.upper()
                # Check references to Detail sheet
                has_detail_ref = 'DETAIL!' in formula_upper
                # Check references to department column (A), budget column (C), variance column (E)
                has_dept_col = bool(re.search(r'DETAIL!A\$?\d', formula_upper, re.IGNORECASE))
                has_budget_col = bool(re.search(r'DETAIL!C\$?\d', formula_upper, re.IGNORECASE))
                has_variance_col = bool(re.search(r'DETAIL!E\$?\d', formula_upper, re.IGNORECASE))

                if has_detail_ref and has_dept_col and has_budget_col and has_variance_col:
                    detail_ref_count += 1
                else:
                    print(f"  INFO: E{row} — Detail ref: {has_detail_ref}, Dept(A): {has_dept_col}, Budget(C): {has_budget_col}, Variance(E): {has_variance_col}")

        if detail_ref_count == 7:
            print(f"PASS: Component 3 — All 7 formulas correctly reference Detail!A, Detail!C, Detail!E (0.30 pts)")
            total_score += 0.30
        elif detail_ref_count > 0:
            partial = round(0.30 * (detail_ref_count / 7), 2)
            print(f"PARTIAL: Component 3 — {detail_ref_count}/7 formulas have correct Detail refs ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No formulas correctly reference Detail sheet columns")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Formula structure implements weighted average pattern (0.20 points)
    # Pattern: SUMPRODUCT(filter * variance * budget) / SUMPRODUCT(filter * budget)
    # The formula should contain a division (/) indicating weighted average
    # and should reference the department cell (D2, D3, etc.) for filtering
    try:
        pattern_count = 0
        for row in range(2, 9):
            cell_val = ws.cell(row=row, column=5).value
            if cell_val and isinstance(cell_val, str):
                formula = cell_val.upper().replace(' ', '')
                # Check for division (weighted average = numerator / denominator)
                has_division = ')/SUMPRODUCT(' in formula or ')/SUMPRODUCT(' in cell_val.replace(' ', '')
                # Check that it references the department cell for filtering (D2, D3, etc.)
                dept_cell_ref = f'D{row}'
                has_dept_filter = dept_cell_ref.upper() in formula

                if has_division and has_dept_filter:
                    pattern_count += 1
                else:
                    print(f"  INFO: E{row} — Division: {has_division}, Dept filter (D{row}): {has_dept_filter}")

        if pattern_count == 7:
            print(f"PASS: Component 4 — All 7 formulas implement weighted average pattern with dept filter (0.20 pts)")
            total_score += 0.20
        elif pattern_count > 0:
            partial = round(0.20 * (pattern_count / 7), 2)
            print(f"PARTIAL: Component 4 — {pattern_count}/7 formulas have correct weighted avg pattern ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No formulas implement the weighted average pattern")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
