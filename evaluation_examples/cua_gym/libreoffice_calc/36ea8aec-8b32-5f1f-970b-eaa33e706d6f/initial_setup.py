"""
Initial Setup: Weighted average variance by department using SUMPRODUCT
Task ID: calc_gg5_034
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Detail Sheet ---
    ws_detail = wb.active
    ws_detail.title = 'Detail'

    headers = ['Department', 'Product Line', 'Budget Amount', 'Actual Amount', 'Variance']
    for col, h in enumerate(headers, 1):
        ws_detail.cell(row=1, column=col, value=h)

    departments = ['Finance', 'Marketing', 'Engineering', 'Operations', 'Sales', 'HR', 'IT']

    product_lines_map = {
        'Finance': ['Audit Services', 'Tax Advisory', 'Financial Planning', 'Risk Management', 'Compliance', 'Treasury'],
        'Marketing': ['Digital Campaigns', 'Brand Strategy', 'Content Marketing', 'Events & Sponsorship', 'Market Research', 'PR & Communications'],
        'Engineering': ['Cloud Infrastructure', 'Mobile Development', 'Data Platform', 'Security Systems', 'DevOps Tooling', 'AI/ML Solutions'],
        'Operations': ['Supply Chain', 'Logistics', 'Quality Assurance', 'Facilities Management', 'Procurement', 'Process Improvement'],
        'Sales': ['Enterprise Accounts', 'SMB Sales', 'Channel Partners', 'Inside Sales', 'Sales Operations', 'Customer Success'],
        'HR': ['Talent Acquisition', 'Learning & Development', 'Compensation & Benefits', 'Employee Relations', 'HR Technology', 'Workforce Analytics'],
        'IT': ['Network Infrastructure', 'Help Desk', 'Software Licensing', 'Cybersecurity', 'Database Administration', 'End User Computing'],
    }

    row_idx = 2
    rows_needed = 300
    rows_per_dept = rows_needed // len(departments)  # ~42 per dept
    extra = rows_needed - rows_per_dept * len(departments)

    for dept_i, dept in enumerate(departments):
        dept_rows = rows_per_dept + (1 if dept_i < extra else 0)
        products = product_lines_map[dept]
        for i in range(dept_rows):
            product = products[i % len(products)]
            budget = round(random.uniform(5000, 150000), 2)
            # Variance ranges from -20% to +15% of budget
            variance_pct = random.uniform(-0.20, 0.15)
            variance = round(budget * variance_pct, 2)
            actual = round(budget + variance, 2)

            ws_detail.cell(row=row_idx, column=1, value=dept)
            ws_detail.cell(row=row_idx, column=2, value=product)
            ws_detail.cell(row=row_idx, column=3, value=budget)
            ws_detail.cell(row=row_idx, column=4, value=actual)
            ws_detail.cell(row=row_idx, column=5, value=variance)
            row_idx += 1

    # Set column widths for readability
    ws_detail.column_dimensions['A'].width = 16
    ws_detail.column_dimensions['B'].width = 24
    ws_detail.column_dimensions['C'].width = 16
    ws_detail.column_dimensions['D'].width = 16
    ws_detail.column_dimensions['E'].width = 14

    # --- Analysis Sheet ---
    ws_analysis = wb.create_sheet('Analysis')

    # Column headers
    ws_analysis.cell(row=1, column=1, value='#')
    ws_analysis.cell(row=1, column=2, value='Department Summary')
    ws_analysis.cell(row=1, column=3, value='Product Lines')
    ws_analysis.cell(row=1, column=4, value='Department')
    # E1 intentionally left empty - agent must add header

    # Department names in D2:D8
    for i, dept in enumerate(departments):
        ws_analysis.cell(row=i + 2, column=1, value=i + 1)
        ws_analysis.cell(row=i + 2, column=2, value=f'{dept} Division')
        ws_analysis.cell(row=i + 2, column=3, value=len(product_lines_map[dept]))
        ws_analysis.cell(row=i + 2, column=4, value=dept)
        # E2:E8 left empty - agent must add SUMPRODUCT formulas

    ws_analysis.column_dimensions['A'].width = 6
    ws_analysis.column_dimensions['B'].width = 22
    ws_analysis.column_dimensions['C'].width = 14
    ws_analysis.column_dimensions['D'].width = 16
    ws_analysis.column_dimensions['E'].width = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
