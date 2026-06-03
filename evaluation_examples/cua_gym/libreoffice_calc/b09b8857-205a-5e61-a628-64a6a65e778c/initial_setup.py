"""
Initial Setup: Apply validation to cell C2 that accepts either a whole number between 1-999 or the text 'N/A'.
Task ID: calc_nrv_074
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Header styling
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Headers
    headers = ['Part', 'Description', 'Quantity']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 14

    # Realistic inventory data
    # Row 2 has C2 EMPTY - that is where the agent must add validation
    data = [
        ['WDG-1024', 'Precision Steel Widget Bracket', None],
        ['FLG-2087', 'Hydraulic Flange Coupler Assembly', 45],
        ['BRG-0553', 'Sealed Ceramic Ball Bearing 6205', 312],
        ['HSG-4471', 'Aluminum Motor Housing Enclosure', 28],
        ['SPR-0089', 'Tempered Compression Spring 50mm', 750],
        ['GKT-3312', 'High-Temp Silicone Gasket Sheet', 'N/A'],
        ['SHF-7720', 'Hardened Drive Shaft 25mm Dia', 16],
        ['CPL-1195', 'Flexible Jaw Coupling Insert', 'N/A'],
        ['VLV-6643', 'Pneumatic Solenoid Valve 24V DC', 93],
        ['PLT-0841', 'CNC Machined Mounting Plate 6061', 204],
        ['NUT-2290', 'Stainless Hex Flange Nut M10', 580],
        ['ROD-5508', 'Chrome Plated Piston Rod 12mm', 67],
        ['BLT-0017', 'Grade 8 Hex Cap Bolt M12x60', 445],
        ['TUB-8834', 'Seamless Stainless Tubing 316L', 'N/A'],
        ['PIN-4462', 'Dowel Pin Hardened 8mm x 30mm', 999],
    ]

    data_font = Font(name="Arial", size=11)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 3:
                cell.alignment = Alignment(horizontal="center")

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
