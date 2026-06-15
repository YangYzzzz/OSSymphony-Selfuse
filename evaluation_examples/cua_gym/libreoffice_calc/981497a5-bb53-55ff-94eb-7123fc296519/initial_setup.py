"""
Initial Setup: Create Transaction_Log spreadsheet with 24 transactions
Task ID: calc_gcv_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transaction_Log"

    # Headers
    headers = ['Transaction ID', 'Customer', 'Date', 'Amount', 'Status']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 24 transactions with realistic data
    # Some refunded rows also have amounts > 1000 (rows with status 'Refunded' and amount > 1000)
    transactions = [
        ['TXN-2025-001', 'Sarah Chen',        '2025-01-05',  245.00, 'Completed'],
        ['TXN-2025-002', 'Marcus Johnson',     '2025-01-08', 1350.00, 'Completed'],
        ['TXN-2025-003', 'Elena Rodriguez',    '2025-01-12',   89.50, 'Pending'],
        ['TXN-2025-004', 'David Kim',          '2025-01-15', 2100.00, 'Refunded'],   # >1000 + Refunded
        ['TXN-2025-005', 'Amara Okafor',       '2025-01-18',  430.00, 'Completed'],
        ['TXN-2025-006', 'James Whitfield',    '2025-01-22', 1875.50, 'Completed'],
        ['TXN-2025-007', 'Priya Sharma',       '2025-01-25',   67.00, 'Refunded'],   # <1000 + Refunded
        ['TXN-2025-008', 'Lucas Fernandez',    '2025-01-28', 3200.00, 'Pending'],
        ['TXN-2025-009', 'Mia Tanaka',         '2025-02-01',  520.00, 'Completed'],
        ['TXN-2025-010', 'Oliver Bennett',     '2025-02-04', 1120.00, 'Refunded'],   # >1000 + Refunded
        ['TXN-2025-011', 'Sophie Martin',      '2025-02-07',  198.75, 'Completed'],
        ['TXN-2025-012', 'Raj Patel',          '2025-02-10', 4500.00, 'Completed'],
        ['TXN-2025-013', 'Hannah Lewis',       '2025-02-13',   50.00, 'Pending'],
        ['TXN-2025-014', 'Chen Wei',           '2025-02-16', 1650.00, 'Completed'],
        ['TXN-2025-015', 'Isabella Costa',     '2025-02-19',  890.00, 'Refunded'],   # <1000 + Refunded
        ['TXN-2025-016', 'Nathan Brooks',      '2025-02-22', 2750.00, 'Pending'],
        ['TXN-2025-017', 'Fatima Al-Hassan',   '2025-02-25',  315.00, 'Completed'],
        ['TXN-2025-018', 'Ryan O\'Connor',     '2025-02-28', 1980.00, 'Refunded'],   # >1000 + Refunded
        ['TXN-2025-019', 'Yuki Nakamura',      '2025-03-03',  175.50, 'Completed'],
        ['TXN-2025-020', 'Grace Adeyemi',      '2025-03-06', 5000.00, 'Completed'],
        ['TXN-2025-021', 'Thomas Muller',      '2025-03-09',  625.00, 'Pending'],
        ['TXN-2025-022', 'Lily Chang',         '2025-03-12', 1450.00, 'Completed'],
        ['TXN-2025-023', 'Andre Silva',        '2025-03-15',  340.00, 'Refunded'],   # <1000 + Refunded
        ['TXN-2025-024', 'Emma Johansson',     '2025-03-18', 3850.00, 'Pending'],
    ]

    for r, row_data in enumerate(transactions, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:  # Amount column
                cell.number_format = '#,##0.00'

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
