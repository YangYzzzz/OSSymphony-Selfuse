"""
Initial Setup: Conference Speaker Proposal Evaluation Spreadsheet
Task ID: calc_grs_063
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proposals"

    # --- Headers ---
    headers = [
        "Submission ID", "Speaker Name", "Talk Title", "Topic Category",
        "Abstract Length", "Proposed Duration"
    ]
    # Reviewer scoring headers: 5 reviewers x 5 criteria
    criteria = ["Relevance", "Novelty", "Speaker Expertise", "Practical Value", "Abstract Quality"]
    for i in range(1, 6):
        for criterion in criteria:
            headers.append(f"R{i} {criterion}")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name="Calibri")
        cell.fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
        cell.font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Speaker Proposal Data (20 rows, unsorted) ---
    proposals = [
        ["SUB-001", "Elena Vasquez", "Building Resilient Microservices with Kubernetes", "Cloud", 287, 45,
         8, 7, 9, 8, 7,  7, 8, 8, 7, 8,  9, 7, 8, 8, 7,  8, 7, 9, 8, 7,  7, 8, 8, 7, 8],
        ["SUB-002", "Marcus Chen", "Zero Trust Architecture in Practice", "Security", 342, 45,
         9, 8, 9, 9, 8,  8, 9, 7, 8, 9,  9, 8, 9, 8, 7,  9, 8, 8, 9, 8,  8, 7, 9, 8, 8],
        ["SUB-003", "Priya Sharma", "Ethical AI: From Theory to Implementation", "AI", 415, 60,
         9, 9, 8, 9, 8,  9, 8, 9, 7, 9,  8, 9, 8, 9, 8,  9, 8, 9, 8, 9,  8, 9, 8, 9, 8],
        ["SUB-004", "James O'Brien", "CI/CD Pipeline Optimization at Scale", "DevOps", 198, 20,
         6, 7, 5, 6, 7,  5, 6, 4, 7, 5,  7, 6, 5, 6, 5,  6, 7, 5, 6, 7,  5, 6, 7, 5, 6],
        ["SUB-005", "Aisha Rahman", "Leading Remote Engineering Teams", "Leadership", 310, 45,
         7, 8, 7, 8, 6,  6, 7, 8, 6, 7,  8, 7, 8, 7, 8,  7, 8, 6, 7, 8,  7, 6, 8, 7, 7],
        ["SUB-006", "Dmitri Petrov", "Real-time Anomaly Detection with LLMs", "AI", 378, 45,
         9, 8, 10, 9, 8,  9, 10, 8, 9, 8,  10, 9, 9, 8, 9,  9, 8, 10, 9, 8,  9, 10, 8, 9, 9],
        ["SUB-007", "Sarah Nakamura", "Serverless Event-Driven Architectures", "Cloud", 256, 45,
         7, 6, 8, 7, 6,  7, 6, 7, 8, 6,  7, 8, 6, 7, 7,  6, 7, 8, 6, 7,  7, 6, 8, 7, 6],
        ["SUB-008", "Carlos Mendez", "Penetration Testing Automation Frameworks", "Security", 290, 45,
         8, 7, 8, 7, 8,  7, 6, 8, 7, 8,  8, 7, 8, 7, 6,  7, 8, 7, 8, 7,  8, 7, 6, 8, 7],
        ["SUB-009", "Hannah Fischer", "GitOps Workflows for Multi-Cloud Deployments", "DevOps", 335, 45,
         7, 8, 7, 6, 8,  8, 7, 6, 8, 7,  7, 8, 7, 6, 8,  8, 7, 6, 7, 8,  7, 6, 8, 7, 7],
        ["SUB-010", "Oluwaseun Adeyemi", "Transformer Models for Code Generation", "AI", 401, 60,
         8, 9, 8, 7, 9,  8, 7, 9, 8, 7,  9, 8, 7, 8, 9,  8, 7, 9, 8, 7,  8, 9, 7, 8, 8],
        ["SUB-011", "Rebecca Thornton", "Building Inclusive Tech Culture", "Leadership", 275, 45,
         6, 5, 7, 6, 5,  5, 7, 6, 5, 6,  6, 5, 7, 6, 5,  5, 6, 7, 5, 6,  6, 5, 7, 6, 5],
        ["SUB-012", "Wei Zhang", "Kubernetes Security Hardening Guide", "Security", 320, 45,
         8, 9, 8, 8, 9,  7, 8, 9, 8, 7,  9, 8, 8, 9, 8,  8, 9, 7, 8, 9,  8, 7, 9, 8, 8],
        ["SUB-013", "Laura Kowalski", "Infrastructure as Code: Terraform Best Practices", "DevOps", 245, 20,
         7, 7, 6, 7, 8,  6, 7, 7, 6, 7,  7, 6, 7, 8, 7,  7, 6, 7, 7, 6,  7, 7, 6, 7, 7],
        ["SUB-014", "Raj Patel", "Computer Vision for Quality Inspection", "AI", 355, 45,
         8, 7, 8, 9, 7,  8, 9, 7, 8, 7,  8, 7, 9, 8, 7,  8, 9, 7, 8, 7,  8, 7, 9, 8, 7],
        ["SUB-015", "Natalie Dubois", "Cloud Cost Optimization Strategies", "Cloud", 230, 20,
         6, 7, 5, 6, 7,  5, 4, 6, 7, 5,  6, 7, 5, 6, 5,  7, 5, 6, 7, 5,  6, 5, 7, 6, 5],
        ["SUB-016", "Thomas Eriksson", "Managing Technical Debt as a Leader", "Leadership", 300, 45,
         7, 6, 8, 7, 6,  7, 8, 6, 7, 8,  7, 6, 8, 7, 6,  6, 7, 8, 7, 6,  7, 8, 6, 7, 7],
        ["SUB-017", "Yuki Tanaka", "Quantum Computing Applications in Cryptography", "Security", 390, 60,
         9, 10, 8, 9, 10,  10, 9, 8, 10, 9,  9, 10, 8, 9, 10,  10, 9, 8, 9, 10,  9, 8, 10, 9, 9],
        ["SUB-018", "Ahmed Hassan", "Monitoring Distributed Systems with OpenTelemetry", "DevOps", 268, 45,
         7, 6, 7, 8, 6,  6, 7, 8, 6, 7,  7, 6, 8, 7, 6,  7, 8, 6, 7, 6,  6, 7, 8, 7, 6],
        ["SUB-019", "Megan Park", "Federated Learning for Privacy-Preserving AI", "AI", 445, 60,
         8, 9, 8, 7, 9,  9, 8, 7, 9, 8,  8, 9, 7, 8, 9,  9, 8, 7, 8, 9,  8, 7, 9, 8, 8],
        ["SUB-020", "Henrik Larsson", "Blockchain Beyond Cryptocurrency", "Other", 312, 45,
         5, 4, 6, 5, 3,  4, 5, 3, 6, 4,  5, 4, 6, 3, 5,  4, 5, 3, 6, 4,  5, 3, 4, 5, 4],
    ]

    for r, row_data in enumerate(proposals, 2):
        # First 6 columns: metadata
        for c, val in enumerate(row_data[:6], 1):
            ws.cell(row=r, column=c, value=val)
        # Score columns (columns 7-31): 25 scores
        scores = row_data[6:]
        for c, val in enumerate(scores, 7):
            ws.cell(row=r, column=c, value=val)

    # --- Data Validation: Topic Category dropdown (column D) ---
    dv_category = DataValidation(
        type="list",
        formula1='"AI,Security,DevOps,Cloud,Leadership,Other"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_category.error = "Invalid category"
    dv_category.errorTitle = "Error"
    dv_category.prompt = "Select topic category"
    dv_category.promptTitle = "Topic Category"
    dv_category.add("D2:D21")
    ws.add_data_validation(dv_category)

    # --- Data Validation: Duration dropdown (column F) ---
    dv_duration = DataValidation(
        type="list",
        formula1='"20,45,60"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_duration.error = "Invalid duration"
    dv_duration.errorTitle = "Error"
    dv_duration.prompt = "Select proposed duration (minutes)"
    dv_duration.promptTitle = "Duration"
    dv_duration.add("F2:F21")
    ws.add_data_validation(dv_duration)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 15  # Submission ID
    ws.column_dimensions["B"].width = 22  # Speaker Name
    ws.column_dimensions["C"].width = 45  # Talk Title
    ws.column_dimensions["D"].width = 18  # Topic Category
    ws.column_dimensions["E"].width = 16  # Abstract Length
    ws.column_dimensions["F"].width = 18  # Proposed Duration

    # Scoring columns narrower
    from openpyxl.utils import get_column_letter
    for col_idx in range(7, 32):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = "A2"

    # --- Row height for header ---
    ws.row_dimensions[1].height = 35

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
