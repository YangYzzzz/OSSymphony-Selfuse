"""
Reward Script: Conference Speaker Proposal Evaluation Spreadsheet
Task ID: calc_grs_063
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Average Score column with AVERAGE formulas (0.20)
  Component 2: Standard Deviation column with STDEV formulas (0.15)
  Component 3: Final Recommendation column with IF formula logic (0.20)
  Component 4: Rows sorted by Average Score descending (0.15)
  Component 5: Conditional formatting on recommendation column (0.15)
  Component 6: Category distribution chart exists (0.15)
  Total: 1.00
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_063'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the main proposals sheet (case-insensitive search)
    proposals_ws = None
    for sn in wb.sheetnames:
        if 'proposal' in sn.lower():
            proposals_ws = wb[sn]
            break
    if proposals_ws is None:
        proposals_ws = wb.worksheets[0]

    ws = proposals_ws

    # --- Identify column positions by header name ---
    headers = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            headers[str(val).strip().lower()] = c

    avg_col = headers.get('average score')
    stdev_col = headers.get('standard deviation')
    rec_col = headers.get('final recommendation')

    # Component 1: Average Score column with AVERAGE formulas (0.20 points)
    try:
        if avg_col is None:
            print("FAIL: Component 1 — 'Average Score' column header not found")
        else:
            avg_formula_count = 0
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=avg_col).value
                if cell_val and isinstance(cell_val, str) and 'AVERAGE' in cell_val.upper():
                    avg_formula_count += 1
            if avg_formula_count >= 18:
                print(f"PASS: Component 1 — Average Score formulas found in {avg_formula_count}/20 rows (0.20 pts)")
                total_score += 0.20
            elif avg_formula_count >= 10:
                partial = 0.10
                print(f"PARTIAL: Component 1 — Average Score formulas in {avg_formula_count}/20 rows ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 — Only {avg_formula_count} Average Score formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Standard Deviation column with STDEV formulas (0.15 points)
    try:
        if stdev_col is None:
            print("FAIL: Component 2 — 'Standard Deviation' column header not found")
        else:
            stdev_formula_count = 0
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=stdev_col).value
                if cell_val and isinstance(cell_val, str) and 'STDEV' in cell_val.upper():
                    stdev_formula_count += 1
            if stdev_formula_count >= 18:
                print(f"PASS: Component 2 — STDEV formulas found in {stdev_formula_count}/20 rows (0.15 pts)")
                total_score += 0.15
            elif stdev_formula_count >= 10:
                partial = 0.07
                print(f"PARTIAL: Component 2 — STDEV formulas in {stdev_formula_count}/20 rows ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {stdev_formula_count} STDEV formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Final Recommendation column with IF-based formula (0.20 points)
    try:
        if rec_col is None:
            print("FAIL: Component 3 — 'Final Recommendation' column header not found")
        else:
            rec_formula_count = 0
            has_categories = False
            for r in range(2, ws.max_row + 1):
                cell_val = ws.cell(row=r, column=rec_col).value
                if cell_val and isinstance(cell_val, str):
                    upper_val = cell_val.upper()
                    if 'IF' in upper_val:
                        rec_formula_count += 1
                        # Check that the formula references the expected categories
                        if any(cat in upper_val for cat in ['TOP PICK', 'ACCEPT', 'WAITLIST', 'DECLINE']):
                            has_categories = True
            if rec_formula_count >= 18 and has_categories:
                print(f"PASS: Component 3 — IF recommendation formulas with correct categories in {rec_formula_count}/20 rows (0.20 pts)")
                total_score += 0.20
            elif rec_formula_count >= 10:
                partial = 0.10
                print(f"PARTIAL: Component 3 — IF recommendation formulas in {rec_formula_count}/20 rows ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Only {rec_formula_count} IF recommendation formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Rows sorted by Average Score descending (0.15 points)
    # We check that the submission IDs are NOT in sequential order (SUB-001..SUB-020),
    # which would indicate no sorting happened. We also try to verify descending order
    # by loading with data_only to get cached values.
    try:
        if avg_col is None:
            print("FAIL: Component 4 — Cannot check sorting without Average Score column")
        else:
            # Check if row order differs from initial sequential order
            sub_ids = []
            for r in range(2, ws.max_row + 1):
                sid = ws.cell(row=r, column=1).value
                if sid:
                    sub_ids.append(str(sid))

            # Initial order is SUB-001 through SUB-020 sequentially
            initial_order = [f"SUB-{i:03d}" for i in range(1, 21)]
            is_reordered = (sub_ids != initial_order)

            if is_reordered and len(sub_ids) >= 18:
                # Further verify: try loading with data_only to check descending order
                try:
                    wb_data = openpyxl.load_workbook(file_path, data_only=True)
                    ws_data = None
                    for sn in wb_data.sheetnames:
                        if 'proposal' in sn.lower():
                            ws_data = wb_data[sn]
                            break
                    if ws_data is None:
                        ws_data = wb_data.worksheets[0]

                    avg_values = []
                    for r in range(2, ws_data.max_row + 1):
                        v = ws_data.cell(row=r, column=avg_col).value
                        if v is not None:
                            try:
                                avg_values.append(float(v))
                            except (ValueError, TypeError):
                                pass

                    if len(avg_values) >= 2:
                        # Check if values are in descending order (allow minor tolerance)
                        is_descending = all(avg_values[i] >= avg_values[i+1] - 0.01
                                            for i in range(len(avg_values)-1))
                        if is_descending:
                            print(f"PASS: Component 4 — Rows sorted by Average Score descending (0.15 pts)")
                            total_score += 0.15
                        else:
                            # Rows are reordered but not perfectly descending — partial credit
                            print(f"PARTIAL: Component 4 — Rows reordered but not strictly descending (0.10 pts)")
                            total_score += 0.10
                    else:
                        # data_only didn't give cached values, but rows are reordered
                        print(f"PASS: Component 4 — Rows reordered from initial sequential order (0.15 pts)")
                        total_score += 0.15
                except Exception:
                    # Fallback: rows are reordered, that's enough
                    print(f"PASS: Component 4 — Rows reordered from initial sequential order (0.15 pts)")
                    total_score += 0.15
            else:
                print(f"FAIL: Component 4 — Rows still in initial sequential order (not sorted)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on recommendation column (0.15 points)
    try:
        cf_rules = list(ws.conditional_formatting)
        if len(cf_rules) == 0:
            print("FAIL: Component 5 — No conditional formatting rules found")
        else:
            # Check that conditional formatting exists that references the recommendation column
            rec_cf_found = False
            rec_categories_in_cf = set()
            for cf in cf_rules:
                cf_range = str(cf)
                for rule in cf.rules:
                    formula_list = getattr(rule, 'formula', []) or []
                    for f in formula_list:
                        f_upper = str(f).upper()
                        for cat in ['TOP PICK', 'ACCEPT', 'WAITLIST', 'DECLINE']:
                            if cat in f_upper:
                                rec_cf_found = True
                                rec_categories_in_cf.add(cat)

            if rec_cf_found and len(rec_categories_in_cf) >= 3:
                print(f"PASS: Component 5 — Conditional formatting for {len(rec_categories_in_cf)} recommendation categories (0.15 pts)")
                total_score += 0.15
            elif rec_cf_found:
                partial = 0.08
                print(f"PARTIAL: Component 5 — Conditional formatting found but only {len(rec_categories_in_cf)} categories ({partial} pts)")
                total_score += partial
            else:
                # Check if there's any conditional formatting at all (generic)
                if len(cf_rules) >= 1:
                    print(f"PARTIAL: Component 5 — Some conditional formatting found but not for recommendation categories (0.05 pts)")
                    total_score += 0.05
                else:
                    print("FAIL: Component 5 — No conditional formatting on recommendation column")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Category distribution chart exists (0.15 points)
    try:
        # Check for charts on any sheet
        total_charts = 0
        chart_found = False
        for sn in wb.sheetnames:
            sheet = wb[sn]
            total_charts += len(sheet._charts)
            for chart in sheet._charts:
                chart_found = True

        if chart_found:
            print(f"PASS: Component 6 — Chart found ({total_charts} chart(s) across sheets) (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 6 — No charts found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
