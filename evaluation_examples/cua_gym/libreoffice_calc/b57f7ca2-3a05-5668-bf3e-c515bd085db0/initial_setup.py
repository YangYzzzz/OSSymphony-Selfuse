"""
Initial Setup: Lab reference database with arXiv IDs but missing DOIs and citation counts
Task ID: osworld_multi_apps_web_references_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user/Desktop'  # File goes on the Desktop
TASK_ID = 'lab_refs'
OUTPUT = f'{WORKDIR}/{TASK_ID}.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TableCellProperties, TextProperties
        from odf import style as odfstyle
        use_odf = True
    except ImportError:
        use_odf = False

    if use_odf:
        # Create using odfpy
        doc = OpenDocumentSpreadsheet()

        # Create a header style (bold)
        header_style = Style(name="HeaderStyle", family="table-cell")
        header_style.addElement(TextProperties(fontweight="bold"))
        doc.styles.addElement(header_style)

        # Create the table (sheet)
        table = Table(name="References")

        # Header row
        headers = ["Title", "arXiv_ID", "DOI", "Citation_Count", "SS_URL"]
        header_row = TableRow()
        for h in headers:
            cell = TableCell(valuetype="string", stylename="HeaderStyle")
            cell.addElement(P(text=h))
            header_row.addElement(cell)
        table.addElement(header_row)

        # Data rows - 7 conference papers with empty DOI, Citation_Count, SS_URL
        papers = [
            ("Attention Is All You Need", "1706.03762"),
            ("BERT: Pre-training of Deep Bidirectional Transformers for Language Understanding", "1810.04805"),
            ("Language Models are Few-Shot Learners", "2005.14165"),
            ("An Image is Worth 16x16 Words: Transformers for Image Recognition at Scale", "2010.11929"),
            ("Deep Residual Learning for Image Recognition", "1512.03385"),
            ("Generative Adversarial Nets", "1406.2661"),
            ("Unsupervised Representation Learning with Deep Convolutional Generative Adversarial Networks", "1511.06434"),
        ]

        for title, arxiv_id in papers:
            row = TableRow()
            # Title
            cell_title = TableCell(valuetype="string")
            cell_title.addElement(P(text=title))
            row.addElement(cell_title)
            # arXiv_ID
            cell_arxiv = TableCell(valuetype="string")
            cell_arxiv.addElement(P(text=arxiv_id))
            row.addElement(cell_arxiv)
            # DOI - empty
            cell_doi = TableCell()
            cell_doi.addElement(P(text=""))
            row.addElement(cell_doi)
            # Citation_Count - empty
            cell_citations = TableCell()
            cell_citations.addElement(P(text=""))
            row.addElement(cell_citations)
            # SS_URL - empty
            cell_url = TableCell()
            cell_url.addElement(P(text=""))
            row.addElement(cell_url)
            table.addElement(row)

        doc.spreadsheet.addElement(table)

        # Ensure Desktop directory exists
        os.makedirs(WORKDIR, exist_ok=True)
        doc.save(OUTPUT)
        print(f'Initial file created: {OUTPUT}')
    else:
        # Fallback: create using openpyxl as .xlsx then rename
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "References"

        headers = ["Title", "arXiv_ID", "DOI", "Citation_Count", "SS_URL"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        papers = [
            ("Attention Is All You Need", "1706.03762"),
            ("BERT: Pre-training of Deep Bidirectional Transformers for Language Understanding", "1810.04805"),
            ("Language Models are Few-Shot Learners", "2005.14165"),
            ("An Image is Worth 16x16 Words: Transformers for Image Recognition at Scale", "2010.11929"),
            ("Deep Residual Learning for Image Recognition", "1512.03385"),
            ("Generative Adversarial Nets", "1406.2661"),
            ("Unsupervised Representation Learning with Deep Convolutional Generative Adversarial Networks", "1511.06434"),
        ]

        for r, (title, arxiv_id) in enumerate(papers, 2):
            ws.cell(row=r, column=1, value=title)
            ws.cell(row=r, column=2, value=arxiv_id)
            # DOI, Citation_Count, SS_URL left empty

        os.makedirs(WORKDIR, exist_ok=True)
        xlsx_path = OUTPUT.replace('.ods', '.xlsx')
        wb.save(xlsx_path)
        print(f'Initial file created (xlsx fallback): {xlsx_path}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    # Also open a browser to Semantic Scholar for the research task
    launch_gui('google-chrome --new-window "https://www.semanticscholar.org/"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
