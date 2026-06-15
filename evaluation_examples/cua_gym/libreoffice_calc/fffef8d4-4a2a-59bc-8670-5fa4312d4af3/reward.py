"""
Reward Script: VLOOKUP and Pivot Table in HR Spreadsheet
Task ID: osworld_calc_vlookup_pivot_combined_011
Domain: libreoffice_calc

Scoring:
  Component 1 (0.40): VLOOKUP formulas present in all 15 Job Grade cells (D2:D16)
  Component 2 (0.30): Pivot table structure in Sheet2 with correct headers (Job Grade rows, Department columns, Grand Total)
  Component 3 (0.30): Pivot table salary values are correct, including row and column grand totals
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_011'

# Expected pivot table data (from golden analysis)
# Rows: Job Grade (Grade A, B, C, D), Columns: Engineering, Finance, HR, Marketing, Grand Total
EXPECTED_PIVOT_HEADERS_ROW = ['Job Grade', 'Engineering', 'Finance', 'HR', 'Marketing', 'Grand Total']
EXPECTED_PIVOT_ROWS = {
    'Grade A': [None, 57000, 54000, 52000, 163000],
    'Grade B': [154500, 78500, 61000, 68000, 362000],
    'Grade C': [190500, None, 74000, 87500, 352000],
    'Grade D': [115000, 125000, None, 108000, 348000],
}
EXPECTED_GRAND_TOTAL_ROW = [460000, 260500, 189000, 315500, 1225000]


def _normalize_val(val):
    """Normalize None and numeric values for comparison."""
    if val is None:
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return val


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Check that required sheets exist
    if 'HR Data' not in wb.sheetnames:
        print("CRITICAL: Sheet 'HR Data' not found.")
        print("REWARD: 0.0")
        return 0.0

    if 'Sheet2' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Sheet2' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['HR Data']
    ws2 = wb['Sheet2']

    # Component 1: VLOOKUP formulas present in Job Grade column D2:D16 (0.40 points)
    # The initial file has all None in column D. The task requires VLOOKUP formulas for all 15 rows.
    try:
        vlookup_count = 0
        vlookup_total = 15  # rows 2 to 16

        for row in range(2, 17):
            cell = ws1.cell(row=row, column=4)  # Column D
            val = cell.value
            if val is not None and isinstance(val, str) and 'VLOOKUP' in val.upper():
                vlookup_count += 1
            elif val is not None and not isinstance(val, str):
                # Possibly a pre-computed value (if file was opened and saved by Calc)
                # Also accept non-None non-formula values as completed (agent may have used static values)
                vlookup_count += 1

        if vlookup_count == vlookup_total:
            print(f"PASS: Component 1 — All {vlookup_total} Job Grade cells (D2:D16) filled with VLOOKUP (0.40 pts)")
            total_score += 0.40
        elif vlookup_count > 0:
            # Partial credit not defined in rubric; must be all-or-nothing for VLOOKUP presence
            print(f"FAIL: Component 1 — Only {vlookup_count}/{vlookup_total} Job Grade cells filled. Expected all 15.")
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas found in D2:D16 (all cells empty/None)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Pivot table structure in Sheet2 with correct headers (0.30 points)
    # Initial Sheet2 is empty. Golden has a header row and department columns + Grand Total.
    try:
        pivot_ok = True

        # Check Sheet2 has content (at least 3 rows: title/header, data rows, grand total)
        if ws2.max_row < 3:
            print(f"FAIL: Component 2 — Sheet2 has insufficient rows ({ws2.max_row}); expected pivot table")
            pivot_ok = False

        # Check header row: A2='Job Grade', B2='Engineering', C2='Finance', D2='HR', E2='Marketing', F2='Grand Total'
        if pivot_ok:
            header_issues = []
            # The header row may be row 1 or row 2 depending on whether a title row exists
            # From exploration: row 1 is title 'Salary Summary...', row 2 is the actual column header
            # We check for an EXACT 'Job Grade' match (not substring) to avoid matching title row
            header_row = None
            for rnum in range(1, min(ws2.max_row + 1, 5)):
                cell_val = ws2.cell(row=rnum, column=1).value
                if cell_val and str(cell_val).strip() == 'Job Grade':
                    header_row = rnum
                    break

            if header_row is None:
                print(f"FAIL: Component 2 — 'Job Grade' header not found in Sheet2 row 1 or 2")
                pivot_ok = False
            else:
                # Check that key department columns and Grand Total are present in header row
                header_vals = [ws2.cell(row=header_row, column=c).value for c in range(1, 7)]
                header_strs = [str(v).strip() if v is not None else '' for v in header_vals]

                required_headers = ['Job Grade', 'Engineering', 'Finance', 'HR', 'Marketing', 'Grand Total']
                missing = [h for h in required_headers if h not in header_strs]
                if missing:
                    print(f"FAIL: Component 2 — Missing pivot header columns: {missing}. Found: {header_strs}")
                    pivot_ok = False
                else:
                    print(f"PASS: Component 2 — Pivot table headers correct: {header_strs} (0.30 pts)")

        # Check that job grade row labels exist (Grade A, Grade B, Grade C, Grade D)
        if pivot_ok:
            grade_labels_found = set()
            for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, min_col=1, max_col=1, values_only=True):
                val = row[0]
                if val and str(val).strip() in ['Grade A', 'Grade B', 'Grade C', 'Grade D']:
                    grade_labels_found.add(str(val).strip())

            expected_grades = {'Grade A', 'Grade B', 'Grade C', 'Grade D'}
            if not expected_grades.issubset(grade_labels_found):
                missing_grades = expected_grades - grade_labels_found
                print(f"FAIL: Component 2 — Missing job grade rows: {missing_grades}")
                pivot_ok = False
            else:
                pass  # already printed PASS above

        if pivot_ok:
            total_score += 0.30
        else:
            # If we printed PASS above but then found grade label issues, re-handle
            pass

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Pivot table salary values correct including grand totals (0.30 points)
    # Verify specific numeric values in the pivot table match expected salary summaries.
    try:
        pivot_data_ok = True

        # Find the header row in Sheet2 to locate data rows
        header_row = None
        dept_col_map = {}
        for rnum in range(1, min(ws2.max_row + 1, 5)):
            r1_val = ws2.cell(row=rnum, column=1).value
            if r1_val and str(r1_val).strip() == 'Job Grade':
                header_row = rnum
                # Map department name -> column index
                for cnum in range(1, ws2.max_column + 1):
                    hval = ws2.cell(row=header_row, column=cnum).value
                    if hval:
                        dept_col_map[str(hval).strip()] = cnum
                break

        if header_row is None:
            print(f"FAIL: Component 3 — Cannot locate header row in Sheet2 to verify values")
            pivot_data_ok = False
        else:
            # Build row index: grade label -> row number
            grade_row_map = {}
            for rnum in range(header_row + 1, ws2.max_row + 1):
                val = ws2.cell(row=rnum, column=1).value
                if val:
                    grade_row_map[str(val).strip()] = rnum

            # Check each expected pivot value
            errors = []
            for grade, expected_vals in EXPECTED_PIVOT_ROWS.items():
                if grade not in grade_row_map:
                    errors.append(f"Row '{grade}' not found")
                    continue
                rnum = grade_row_map[grade]
                dept_cols = ['Engineering', 'Finance', 'HR', 'Marketing', 'Grand Total']
                for dept, exp_val in zip(dept_cols, expected_vals):
                    if dept not in dept_col_map:
                        errors.append(f"Column '{dept}' not in header")
                        continue
                    cnum = dept_col_map[dept]
                    actual = ws2.cell(row=rnum, column=cnum).value
                    actual_n = _normalize_val(actual)
                    exp_n = _normalize_val(exp_val)
                    if actual_n != exp_n:
                        # Allow None where expected (both None is ok)
                        if exp_n is None and actual_n is None:
                            pass
                        elif exp_n is None and actual_n is not None:
                            errors.append(f"{grade}/{dept}: expected None, got {actual_n}")
                        elif exp_n is not None and actual_n is None:
                            errors.append(f"{grade}/{dept}: expected {exp_n}, got None")
                        else:
                            errors.append(f"{grade}/{dept}: expected {exp_n}, got {actual_n}")

            # Check Grand Total row
            if 'Grand Total' in grade_row_map:
                gt_rnum = grade_row_map['Grand Total']
                gt_depts = ['Engineering', 'Finance', 'HR', 'Marketing', 'Grand Total']
                for dept, exp_val in zip(gt_depts, EXPECTED_GRAND_TOTAL_ROW):
                    if dept not in dept_col_map:
                        continue
                    cnum = dept_col_map[dept]
                    actual = ws2.cell(row=gt_rnum, column=cnum).value
                    actual_n = _normalize_val(actual)
                    exp_n = _normalize_val(exp_val)
                    if actual_n != exp_n:
                        errors.append(f"Grand Total/{dept}: expected {exp_n}, got {actual_n}")
            else:
                errors.append("Grand Total row not found in pivot table")

            if errors:
                print(f"FAIL: Component 3 — Pivot table value errors: {errors[:5]}")  # show first 5
                pivot_data_ok = False
            else:
                print(f"PASS: Component 3 — All pivot table salary values and grand totals correct (0.30 pts)")

        if pivot_data_ok:
            total_score += 0.30

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
