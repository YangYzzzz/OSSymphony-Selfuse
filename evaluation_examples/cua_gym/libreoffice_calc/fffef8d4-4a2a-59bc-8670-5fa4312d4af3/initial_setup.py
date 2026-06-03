"""
Initial Setup: VLOOKUP + Pivot Table Combined Task
Task ID: osworld_calc_vlookup_pivot_combined_011
Domain: libreoffice_calc

Creates an HR data spreadsheet with:
- Sheet1: Employee data with Employee ID, Department, Job Code,
          Job Grade (EMPTY - agent must fill via VLOOKUP), Annual Salary
- Columns G-H in Sheet1: Job Code -> Job Grade reference table
- NO pivot table in Sheet2 (agent must create it)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: HR Data ---
    ws1 = wb.active
    ws1.title = 'HR Data'

    # Header row styling
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Main data headers (columns A-E)
    main_headers = ['Employee ID', 'Department', 'Job Code', 'Job Grade', 'Annual Salary']
    for col, h in enumerate(main_headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    # Reference table headers (columns G-H)
    ref_headers = ['Job Code', 'Job Grade']
    for col, h in enumerate(ref_headers, 7):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border

    # Job Code -> Job Grade reference table (G2:H7)
    job_grade_lookup = [
        ('JC-101', 'Grade A'),
        ('JC-102', 'Grade B'),
        ('JC-103', 'Grade B'),
        ('JC-104', 'Grade C'),
        ('JC-105', 'Grade C'),
        ('JC-106', 'Grade D'),
    ]
    for r, (jc, jg) in enumerate(job_grade_lookup, 2):
        ws1.cell(row=r, column=7, value=jc).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        ws1.cell(row=r, column=8, value=jg).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Employee data (columns A-E, rows 2-16)
    # Job Grade column (D) is intentionally EMPTY - agent must fill via VLOOKUP
    employee_data = [
        ('EMP-001', 'Engineering',  'JC-104', None, 92500),
        ('EMP-002', 'Marketing',    'JC-102', None, 68000),
        ('EMP-003', 'Engineering',  'JC-106', None, 115000),
        ('EMP-004', 'HR',           'JC-101', None, 54000),
        ('EMP-005', 'Finance',      'JC-103', None, 78500),
        ('EMP-006', 'Engineering',  'JC-105', None, 98000),
        ('EMP-007', 'Marketing',    'JC-101', None, 52000),
        ('EMP-008', 'Finance',      'JC-106', None, 125000),
        ('EMP-009', 'HR',           'JC-102', None, 61000),
        ('EMP-010', 'Engineering',  'JC-103', None, 83000),
        ('EMP-011', 'Marketing',    'JC-104', None, 87500),
        ('EMP-012', 'Finance',      'JC-101', None, 57000),
        ('EMP-013', 'HR',           'JC-105', None, 74000),
        ('EMP-014', 'Engineering',  'JC-102', None, 71500),
        ('EMP-015', 'Marketing',    'JC-106', None, 108000),
    ]

    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r, row_data in enumerate(employee_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = data_border
            if c == 5 and val is not None:
                cell.number_format = '$#,##0'

    # Column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 4   # spacer
    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 12

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet2: (empty - agent must create pivot table here) ---
    ws2 = wb.create_sheet('Sheet2')
    # Leave Sheet2 empty; agent will create pivot table here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
