"""
Initial Setup: Large config spreadsheet with 3-row header and identifier columns
Task ID: calc_ggf_036
Domain: libreoffice_calc

Creates a spreadsheet with a Config sheet containing:
- Row 1: Report title
- Row 2: Filter criteria
- Row 3: Column labels
- Rows 4-300: Configuration parameter data
- Columns A-B: identifiers (Config ID, Parameter Name)
- Columns C-Z: various config values
No freeze panes, no background colors applied.
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Config"

    # --- Row 1: Report title (merged across many columns) ---
    ws.merge_cells("A1:Z1")
    ws["A1"] = "Infrastructure Configuration Report — Q1 2026"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 2: Filter criteria ---
    filter_labels = [
        "Region:", "All", "Environment:", "Production + Staging",
        "Status:", "Active", "Last Reviewed:", "2026-03-28",
        "Owner:", "Platform Engineering", "Compliance:", "SOC2",
        "Priority:", "P1-P3", "Category:", "All",
        "Audit Period:", "Q1 2026", "Approver:", "J. Martinez",
        "Version:", "3.2.1", "Export Date:", "2026-03-30",
    ]
    for col_idx, label in enumerate(filter_labels, 1):
        c = ws.cell(row=2, column=col_idx, value=label)
        if col_idx % 2 == 1:
            c.font = Font(bold=True, size=10)
        else:
            c.font = Font(size=10)

    # --- Row 3: Column labels ---
    headers = [
        "Config ID", "Parameter Name", "Current Value", "Default Value",
        "Data Type", "Environment", "Service", "Region",
        "Owner", "Last Modified", "Status", "Priority",
        "Description", "Validation Rule", "Dependencies",
        "Rollback Value", "Change Ticket", "Approved By",
        "Compliance Tag", "Audit Notes", "TTL (days)",
        "Encrypted", "Source", "Version", "Deprecated", "Comments"
    ]
    for col_idx, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col_idx, value=h)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center", wrap_text=True)

    # --- Rows 4-300: Configuration data ---
    services = [
        "auth-service", "api-gateway", "user-mgmt", "payment-processor",
        "notification-engine", "data-pipeline", "cache-layer", "search-index",
        "logging-service", "cdn-proxy", "scheduler", "analytics-engine",
        "rate-limiter", "config-server", "vault-agent", "dns-resolver",
        "load-balancer", "message-queue", "db-proxy", "monitoring-agent"
    ]
    param_names = [
        "max_connections", "timeout_ms", "retry_count", "batch_size",
        "log_level", "cache_ttl_sec", "rate_limit_rps", "thread_pool_size",
        "heap_memory_mb", "gc_interval_sec", "tls_version", "cipher_suite",
        "idle_timeout_ms", "max_payload_kb", "compression_enabled",
        "circuit_breaker_threshold", "health_check_interval", "dns_ttl",
        "connection_pool_min", "connection_pool_max"
    ]
    environments = ["production", "staging", "development", "qa", "canary"]
    regions = ["us-east-1", "us-west-2", "eu-west-1", "ap-southeast-1", "ap-northeast-1"]
    statuses = ["Active", "Pending Review", "Deprecated", "Under Audit", "Approved"]
    priorities = ["P1 - Critical", "P2 - High", "P3 - Medium", "P4 - Low"]
    data_types = ["integer", "string", "boolean", "float", "enum", "duration"]
    owners = [
        "Sarah Chen", "Marcus Johnson", "Priya Patel", "David Kim",
        "Elena Rodriguez", "James Wright", "Aisha Mohammed", "Wei Zhang",
        "Carlos Mendez", "Rachel Green", "Tom Baker", "Nina Kowalski"
    ]
    approvers = [
        "J. Martinez", "R. Thompson", "A. Singh", "M. O'Brien",
        "K. Yamamoto", "L. Schmidt"
    ]
    compliance_tags = ["SOC2", "HIPAA", "PCI-DSS", "GDPR", "ISO27001", "FedRAMP", "None"]
    sources = ["terraform", "consul", "vault", "env-file", "k8s-configmap", "manual"]
    validation_rules = [
        "range(1,10000)", "enum(DEBUG,INFO,WARN,ERROR)", "boolean",
        "range(64,8192)", "regex(TLSv1\\.[2-3])", "range(0,100)",
        "string_max(256)", "ip_address", "port_range(1024,65535)", "duration_ms"
    ]

    for row_idx in range(4, 301):
        cfg_id = f"CFG-{row_idx - 3:04d}"
        svc = random.choice(services)
        param = random.choice(param_names)
        env_name = random.choice(environments)
        region = random.choice(regions)

        # Generate plausible values based on param type
        if "timeout" in param or "ttl" in param or "interval" in param:
            cur_val = str(random.choice([100, 250, 500, 1000, 3000, 5000, 10000, 30000]))
            def_val = "5000"
            dtype = "integer"
        elif "count" in param or "size" in param or "pool" in param:
            cur_val = str(random.randint(1, 256))
            def_val = str(random.choice([4, 8, 16, 32]))
            dtype = "integer"
        elif "enabled" in param:
            cur_val = random.choice(["true", "false"])
            def_val = "false"
            dtype = "boolean"
        elif "level" in param:
            cur_val = random.choice(["DEBUG", "INFO", "WARN", "ERROR"])
            def_val = "INFO"
            dtype = "enum"
        elif "version" in param or "suite" in param:
            cur_val = random.choice(["TLSv1.2", "TLSv1.3", "AES-256-GCM"])
            def_val = "TLSv1.2"
            dtype = "string"
        elif "memory" in param or "payload" in param:
            cur_val = str(random.choice([256, 512, 1024, 2048, 4096]))
            def_val = "512"
            dtype = "integer"
        elif "limit" in param or "threshold" in param:
            cur_val = str(random.randint(10, 10000))
            def_val = "1000"
            dtype = "integer"
        else:
            cur_val = str(random.randint(1, 9999))
            def_val = str(random.randint(1, 100))
            dtype = random.choice(data_types)

        owner = random.choice(owners)
        year = 2026
        month = random.randint(1, 3)
        day = random.randint(1, 28)
        last_mod = f"{year}-{month:02d}-{day:02d}"
        status = random.choice(statuses)
        priority = random.choice(priorities)
        desc = f"Controls {param.replace('_', ' ')} for {svc}"
        val_rule = random.choice(validation_rules)
        deps = random.choice(["None", f"{random.choice(services)}.{random.choice(param_names)}",
                               f"CFG-{random.randint(1, 200):04d}"])
        rollback = def_val
        ticket = f"CHG-{random.randint(10000, 99999)}"
        approver = random.choice(approvers)
        comp_tag = random.choice(compliance_tags)
        audit = random.choice(["Reviewed", "Pending", "N/A", "Flagged", ""])
        ttl = random.choice([30, 60, 90, 180, 365, 0])
        encrypted = random.choice(["Yes", "No"])
        source = random.choice(sources)
        version = f"{random.randint(1,5)}.{random.randint(0,9)}.{random.randint(0,9)}"
        deprecated = random.choice(["No", "No", "No", "Yes"])
        comment = random.choice([
            "", "Tuned for peak traffic", "Requires restart after change",
            "Inherited from base config", "Override for compliance",
            "Temporary — revert after migration", "Approved in Q4 review",
            "Performance-critical setting"
        ])

        row_data = [
            cfg_id, f"{svc}.{param}", cur_val, def_val,
            dtype, env_name, svc, region,
            owner, last_mod, status, priority,
            desc, val_rule, deps,
            rollback, ticket, approver,
            comp_tag, audit, ttl,
            encrypted, source, version, deprecated, comment
        ]
        for c, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=c, value=val)

    # Set column widths for readability
    col_widths = {
        'A': 12, 'B': 35, 'C': 16, 'D': 14, 'E': 12, 'F': 14,
        'G': 20, 'H': 18, 'I': 18, 'J': 14, 'K': 16, 'L': 16,
        'M': 40, 'N': 22, 'O': 30, 'P': 14, 'Q': 14, 'R': 16,
        'S': 14, 'T': 14, 'U': 10, 'V': 10, 'W': 16, 'X': 10,
        'Y': 12, 'Z': 30
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 24

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
