"""
Reward Script: Freeze panes and apply matching background colors
Task ID: calc_ggf_036
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Freeze panes at C4 (rows 1-3 + cols A-B)
  Component 2 (0.35): Light yellow background on rows 2-3, cols C-Z
  Component 3 (0.30): Light blue background on cols A-B, rows 1-300
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_036'

# Accepted color codes (case-insensitive comparison)
LIGHT_YELLOW = 'FFFFFF00'
LIGHT_BLUE = 'FFADD8E6'

# We also accept some near-variants for light yellow / light blue
YELLOW_VARIANTS = {'FFFFFF00', 'FFFFFFCC', 'FFFFFFE0', 'FFFFF2CC', 'FFFFFF99'}
BLUE_VARIANTS = {'FFADD8E6', 'FFBDD7EE', 'FFD6ECFF', 'FF9BC2E6', 'FFDCE6F1'}


def get_fill_rgb(cell):
    """Get the fill foreground color RGB as uppercase string, or None."""
    try:
        if cell.fill.fill_type == 'solid' or cell.fill.patternType == 'solid':
            rgb = cell.fill.fgColor.rgb
            if rgb and rgb != '00000000':
                return str(rgb).upper()
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Config' sheet must exist
    if 'Config' not in wb.sheetnames:
        print("CRITICAL: 'Config' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Config']

    # =========================================================================
    # Component 1: Freeze panes at C4 (0.35 points)
    # This freezes rows 1-3 and columns A-B simultaneously.
    # Initial state has freeze_panes=None, so this only passes on golden.
    # =========================================================================
    try:
        fp = ws.freeze_panes
        if fp == 'C4':
            print(f"PASS: Component 1 — Freeze panes set to C4 (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Expected freeze_panes='C4', found '{fp}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Light yellow background on rows 2-3, columns C-Z (0.35 pts)
    # Row 1 is merged A1:Z1, so C1:Z1 are MergedCells that cannot hold fills.
    # We check rows 2-3, columns C(3) through Z(26).
    # Sample cells to verify: check a representative set.
    # Initial state has no fills, so this only passes on golden.
    # =========================================================================
    try:
        yellow_pass = 0
        yellow_total = 0
        # Check all cells in rows 2-3, columns C-Z
        sample_cols = [3, 4, 5, 10, 15, 20, 25, 26]  # C, D, E, J, O, T, Y, Z
        sample_rows = [2, 3]
        for r in sample_rows:
            for c in sample_cols:
                yellow_total += 1
                cell = ws.cell(row=r, column=c)
                rgb = get_fill_rgb(cell)
                if rgb and rgb in YELLOW_VARIANTS:
                    yellow_pass += 1

        if yellow_total > 0:
            yellow_ratio = yellow_pass / yellow_total
        else:
            yellow_ratio = 0.0

        if yellow_ratio >= 0.8:
            print(f"PASS: Component 2 — Light yellow on rows 2-3, cols C-Z ({yellow_pass}/{yellow_total} cells) (0.35 pts)")
            total_score += 0.35
        elif yellow_ratio >= 0.5:
            partial = round(0.35 * yellow_ratio, 2)
            print(f"PARTIAL: Component 2 — Yellow fill on {yellow_pass}/{yellow_total} sampled cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Yellow fill on only {yellow_pass}/{yellow_total} sampled cells")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Light blue background on columns A-B, rows 1-300 (0.30 pts)
    # Row 1: A1 is the merged top-left, so it should have fill.
    #         B1 is a MergedCell, so we skip it.
    # Rows 2-300: both A and B should have light blue.
    # Initial state has no fills, so this only passes on golden.
    # =========================================================================
    try:
        blue_pass = 0
        blue_total = 0

        # Check A1 (merged cell top-left, should have blue fill)
        rgb_a1 = get_fill_rgb(ws['A1'])
        blue_total += 1
        if rgb_a1 and rgb_a1 in BLUE_VARIANTS:
            blue_pass += 1

        # Sample rows for columns A and B (skip B1 as it's merged)
        sample_rows_blue = [2, 3, 4, 5, 10, 50, 100, 150, 200, 250, 299, 300]
        for r in sample_rows_blue:
            for c in [1, 2]:  # A, B
                blue_total += 1
                cell = ws.cell(row=r, column=c)
                if isinstance(cell, MergedCell):
                    continue  # skip merged cells, don't count against
                rgb = get_fill_rgb(cell)
                if rgb and rgb in BLUE_VARIANTS:
                    blue_pass += 1

        if blue_total > 0:
            blue_ratio = blue_pass / blue_total
        else:
            blue_ratio = 0.0

        if blue_ratio >= 0.8:
            print(f"PASS: Component 3 — Light blue on cols A-B, rows 1-300 ({blue_pass}/{blue_total} cells) (0.30 pts)")
            total_score += 0.30
        elif blue_ratio >= 0.5:
            partial = round(0.30 * blue_ratio, 2)
            print(f"PARTIAL: Component 3 — Blue fill on {blue_pass}/{blue_total} sampled cells ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Blue fill on only {blue_pass}/{blue_total} sampled cells")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
