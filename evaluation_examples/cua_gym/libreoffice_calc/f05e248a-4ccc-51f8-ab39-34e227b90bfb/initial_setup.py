"""
Initial Setup: Build a quarterly summary table in Sheet2 aggregating revenue and expenses
Task ID: osworld_calc_sheet2_summary_table_002
Domain: libreoffice_calc

Creates a spreadsheet with:
  - Sheet1 "Transactions": Detailed financial transaction data (Date, Type, Category, Amount)
  - Sheet2 "Summary": Empty table - no aggregation data yet (agent must add it)
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions ---
    ws1 = wb.active
    ws1.title = 'Transactions'

    # Headers
    headers = ['Date', 'Type', 'Category', 'Amount']
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Realistic transaction data spread across Q1-Q4 of 2024
    # Dates as strings for simplicity (LibreOffice will parse them)
    transactions = [
        # Q1: Jan-Mar
        ('2024-01-08', 'Revenue',  'Product Sales',       42500.00),
        ('2024-01-15', 'Expense',  'Office Supplies',      1250.00),
        ('2024-01-22', 'Revenue',  'Service Contracts',   18300.00),
        ('2024-02-05', 'Expense',  'Salaries',            38000.00),
        ('2024-02-14', 'Revenue',  'Product Sales',       31750.00),
        ('2024-02-20', 'Expense',  'Marketing',            5400.00),
        ('2024-03-03', 'Revenue',  'Consulting',          22000.00),
        ('2024-03-18', 'Expense',  'Rent',                 8500.00),
        ('2024-03-28', 'Revenue',  'Service Contracts',   15600.00),
        # Q2: Apr-Jun
        ('2024-04-05', 'Revenue',  'Product Sales',       48200.00),
        ('2024-04-12', 'Expense',  'Salaries',            38000.00),
        ('2024-04-25', 'Expense',  'Utilities',            2100.00),
        ('2024-05-07', 'Revenue',  'Consulting',          27500.00),
        ('2024-05-16', 'Expense',  'Office Supplies',      980.00),
        ('2024-05-29', 'Revenue',  'Product Sales',       36400.00),
        ('2024-06-04', 'Expense',  'Marketing',            7200.00),
        ('2024-06-18', 'Revenue',  'Service Contracts',   19800.00),
        ('2024-06-30', 'Expense',  'Rent',                 8500.00),
        # Q3: Jul-Sep
        ('2024-07-08', 'Revenue',  'Product Sales',       52100.00),
        ('2024-07-15', 'Expense',  'Salaries',            39500.00),
        ('2024-07-22', 'Expense',  'Equipment',           12000.00),
        ('2024-08-06', 'Revenue',  'Consulting',          31000.00),
        ('2024-08-19', 'Expense',  'Marketing',            6800.00),
        ('2024-08-27', 'Revenue',  'Product Sales',       44700.00),
        ('2024-09-10', 'Expense',  'Rent',                 8500.00),
        ('2024-09-24', 'Revenue',  'Service Contracts',   21300.00),
        # Q4: Oct-Dec
        ('2024-10-03', 'Revenue',  'Product Sales',       58900.00),
        ('2024-10-14', 'Expense',  'Salaries',            39500.00),
        ('2024-10-28', 'Expense',  'Marketing',            9500.00),
        ('2024-11-06', 'Revenue',  'Consulting',          35200.00),
        ('2024-11-19', 'Expense',  'Office Supplies',      1600.00),
        ('2024-11-25', 'Revenue',  'Product Sales',       49300.00),
        ('2024-12-09', 'Expense',  'Rent',                 8500.00),
        ('2024-12-17', 'Revenue',  'Service Contracts',   24100.00),
        ('2024-12-28', 'Expense',  'Utilities',            2400.00),
    ]

    for r, row_data in enumerate(transactions, 2):
        ws1.cell(row=r, column=1, value=row_data[0])
        ws1.cell(row=r, column=2, value=row_data[1])
        ws1.cell(row=r, column=3, value=row_data[2])
        ws1.cell(row=r, column=4, value=row_data[3])

    # Column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 14

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Summary (empty - agent must fill this in) ---
    ws2 = wb.create_sheet('Summary')

    # Only add the sheet title - leave summary table empty for agent to create
    ws2.cell(row=1, column=1, value='Quarterly Financial Summary')
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
