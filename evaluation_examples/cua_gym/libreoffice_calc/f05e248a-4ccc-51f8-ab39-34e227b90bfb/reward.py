"""
Reward Script: Build a quarterly summary table in Sheet2 (Summary)
Task ID: osworld_calc_sheet2_summary_table_002
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.4 pts): Summary sheet has header row with correct column labels
                          (Quarter, Total Revenue, Total Expenses) in row 2
  Component 2 (0.3 pts): Summary sheet has all 4 quarter labels Q1-Q4 in column A
  Component 3 (0.3 pts): Revenue and Expense aggregation formulas/values present
                          for all 4 quarters in columns B and C

Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_sheet2_summary_table_002'

# Expected quarter labels and their canonical row positions
QUARTER_LABELS = ['Q1', 'Q2', 'Q3', 'Q4']

# Pre-calculated expected aggregation values from Transactions sheet
# These are the correct totals based on the transaction data
EXPECTED_REVENUES = {
    'Q1': 130150,   # Jan-Mar 2024
    'Q2': 131900,   # Apr-Jun 2024
    'Q3': 149100,   # Jul-Sep 2024
    'Q4': 167500,   # Oct-Dec 2024
}
EXPECTED_EXPENSES = {
    'Q1': 53150,    # Jan-Mar 2024
    'Q2': 56780,    # Apr-Jun 2024
    'Q3': 66800,    # Jul-Sep 2024
    'Q4': 61500,    # Oct-Dec 2024
}


def has_aggregation_formula_or_value(cell_value, quarter, agg_type):
    """
    Check if a cell value represents a valid revenue/expense aggregation.
    Accepts either a SUMIF/SUMIFS formula string or the pre-calculated numeric value.
    """
    if cell_value is None:
        return False

    val_str = str(cell_value).upper().strip()

    # Accept any SUMIF or SUMIFS formula referencing Transactions sheet
    if 'SUMIF' in val_str and 'TRANSACTION' in val_str:
        return True

    # Accept numeric values that match expected totals
    expected = EXPECTED_REVENUES[quarter] if agg_type == 'revenue' else EXPECTED_EXPENSES[quarter]
    try:
        numeric = float(cell_value)
        if abs(numeric - expected) <= 1.0:
            return True
    except (TypeError, ValueError):
        pass

    return False


def verify_task(file_path):
    """
    Verify that Summary sheet has a quarterly summary table
    with Quarter, Total Revenue, Total Expenses columns for Q1-Q4.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found in workbook")
        print(f"REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: Header row with correct column labels (0.4 points)
    # The header row should have Quarter, Total Revenue, Total Expenses in some row
    # We search rows 1-5 for a header row pattern
    try:
        header_row_found = False
        quarter_col = None
        revenue_col = None
        expense_col = None
        header_row_idx = None

        for row_idx in range(1, min(6, ws.max_row + 1)):
            row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            row_strings = [str(v).strip().lower() if v is not None else '' for v in row_values]

            # Look for 'quarter' label
            q_col = None
            r_col = None
            e_col = None
            for i, v in enumerate(row_strings):
                if 'quarter' in v:
                    q_col = i + 1
                if 'revenue' in v:
                    r_col = i + 1
                if 'expense' in v:
                    e_col = i + 1

            if q_col and r_col and e_col:
                header_row_found = True
                quarter_col = q_col
                revenue_col = r_col
                expense_col = e_col
                header_row_idx = row_idx
                break

        if header_row_found:
            print(f"PASS: Component 1 — Header row found at row {header_row_idx} "
                  f"(Quarter col={quarter_col}, Revenue col={revenue_col}, Expense col={expense_col}) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — No header row with 'Quarter', 'Total Revenue', 'Total Expenses' found in rows 1-5")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        header_row_found = False
        quarter_col = None
        revenue_col = None
        expense_col = None
        header_row_idx = None

    # Component 2: All 4 quarter labels Q1-Q4 present in column A (0.3 points)
    try:
        if not header_row_found or quarter_col is None:
            print("FAIL: Component 2 — Skipped because header row not found")
        else:
            # Collect all values in the quarter column below the header
            quarter_values_found = set()
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=quarter_col).value
                if cell_val is not None:
                    quarter_values_found.add(str(cell_val).strip().upper())

            missing_quarters = [q for q in QUARTER_LABELS if q not in quarter_values_found]

            if not missing_quarters:
                print(f"PASS: Component 2 — All 4 quarter labels (Q1, Q2, Q3, Q4) found in column {quarter_col} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Missing quarter labels: {missing_quarters}. "
                      f"Found: {sorted(quarter_values_found)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Revenue and Expense aggregation values/formulas for all 4 quarters (0.3 points)
    try:
        if not header_row_found or revenue_col is None or expense_col is None:
            print("FAIL: Component 3 — Skipped because header row not found")
        else:
            # Map quarter labels to their row positions
            quarter_rows = {}
            for row_idx in range(header_row_idx + 1, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=quarter_col).value
                if cell_val is not None:
                    q_label = str(cell_val).strip().upper()
                    if q_label in QUARTER_LABELS:
                        quarter_rows[q_label] = row_idx

            agg_checks_passed = 0
            agg_checks_total = 0

            for q in QUARTER_LABELS:
                if q not in quarter_rows:
                    print(f"  SKIP: Quarter {q} row not found")
                    continue

                row_idx = quarter_rows[q]

                # Check revenue cell
                rev_cell = ws.cell(row=row_idx, column=revenue_col).value
                exp_cell = ws.cell(row=row_idx, column=expense_col).value

                agg_checks_total += 2

                if has_aggregation_formula_or_value(rev_cell, q, 'revenue'):
                    agg_checks_passed += 1
                    print(f"  PASS: {q} revenue cell has valid formula/value: {str(rev_cell)[:60]}")
                else:
                    print(f"  FAIL: {q} revenue cell invalid (found: {repr(rev_cell)}, "
                          f"expected SUMIF formula or ~{EXPECTED_REVENUES[q]})")

                if has_aggregation_formula_or_value(exp_cell, q, 'expense'):
                    agg_checks_passed += 1
                    print(f"  PASS: {q} expense cell has valid formula/value: {str(exp_cell)[:60]}")
                else:
                    print(f"  FAIL: {q} expense cell invalid (found: {repr(exp_cell)}, "
                          f"expected SUMIF formula or ~{EXPECTED_EXPENSES[q]})")

            if agg_checks_total > 0 and agg_checks_passed == agg_checks_total:
                print(f"PASS: Component 3 — All {agg_checks_total} revenue/expense aggregation cells "
                      f"verified ({agg_checks_passed}/{agg_checks_total}) (0.3 pts)")
                total_score += 0.3
            elif agg_checks_total > 0 and agg_checks_passed > 0:
                partial = 0.3 * (agg_checks_passed / agg_checks_total)
                print(f"PARTIAL: Component 3 — {agg_checks_passed}/{agg_checks_total} aggregation cells valid "
                      f"(+{partial:.2f} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — No valid revenue/expense aggregation cells found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
