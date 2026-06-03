"""
Initial Setup: Set row height and enable word wrap for header row
Task ID: calc_gfl_081
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Activity Log'

    # Headers (row 1) - long multi-word headers that overflow at default height
    headers = [
        'Transaction Reference Number',
        'Transaction Date',
        'Account Holder',
        'Debit Amount',
        'Credit Amount',
        'Running Balance',
        'Branch Code',
        'Authorized By',
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        # Do NOT set wrap_text - that's the task

    # Set column widths so data looks reasonable (but header text still overflows)
    col_widths = [12, 14, 16, 13, 13, 14, 11, 14]
    for i, width in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = width

    # Row 1 at default height - do NOT set custom height
    # (default is ~0.45cm / ~12.75 points)

    # 27 rows of realistic banking transaction data
    data = [
        ['TXN-2025-00143', '2025-01-03', 'Sarah Chen', 2500.00, None, 47500.00, 'BR-101', 'James Liu'],
        ['TXN-2025-00144', '2025-01-03', 'Marcus Johnson', None, 8500.00, 56000.00, 'BR-204', 'Priya Sharma'],
        ['TXN-2025-00145', '2025-01-04', 'Elena Rodriguez', 1200.00, None, 54800.00, 'BR-101', 'James Liu'],
        ['TXN-2025-00146', '2025-01-05', 'David Kim', 3750.00, None, 51050.00, 'BR-307', 'Robert Chang'],
        ['TXN-2025-00147', '2025-01-06', 'Aisha Patel', None, 15000.00, 66050.00, 'BR-204', 'Priya Sharma'],
        ['TXN-2025-00148', '2025-01-07', 'Thomas Weber', 890.50, None, 65159.50, 'BR-101', 'James Liu'],
        ['TXN-2025-00149', '2025-01-08', 'Fatima Al-Said', 4200.00, None, 60959.50, 'BR-502', 'Nina Petrov'],
        ['TXN-2025-00150', '2025-01-09', 'Carlos Mendez', None, 6300.00, 67259.50, 'BR-307', 'Robert Chang'],
        ['TXN-2025-00151', '2025-01-10', 'Yuki Tanaka', 1575.25, None, 65684.25, 'BR-204', 'Priya Sharma'],
        ['TXN-2025-00152', '2025-01-11', 'Rachel Green', 9800.00, None, 55884.25, 'BR-101', 'James Liu'],
        ['TXN-2025-00153', '2025-01-13', 'Omar Hassan', None, 22000.00, 77884.25, 'BR-502', 'Nina Petrov'],
        ['TXN-2025-00154', '2025-01-14', 'Lisa Chang', 3100.00, None, 74784.25, 'BR-307', 'Robert Chang'],
        ['TXN-2025-00155', '2025-01-15', 'Benjamin Okafor', 670.00, None, 74114.25, 'BR-101', 'James Liu'],
        ['TXN-2025-00156', '2025-01-16', 'Sophie Martin', None, 4500.00, 78614.25, 'BR-204', 'Priya Sharma'],
        ['TXN-2025-00157', '2025-01-17', 'Raj Krishnamurthy', 12500.00, None, 66114.25, 'BR-502', 'Nina Petrov'],
        ['TXN-2025-00158', '2025-01-20', 'Hannah Mueller', 2350.00, None, 63764.25, 'BR-307', 'Robert Chang'],
        ['TXN-2025-00159', '2025-01-21', 'Antonio Silva', None, 9750.00, 73514.25, 'BR-101', 'James Liu'],
        ['TXN-2025-00160', '2025-01-22', 'Mei Lin Wang', 5680.00, None, 67834.25, 'BR-204', 'Priya Sharma'],
        ['TXN-2025-00161', '2025-01-23', 'Isaac Newton', 430.75, None, 67403.50, 'BR-502', 'Nina Petrov'],
        ['TXN-2025-00162', '2025-01-24', 'Grace Adeyemi', None, 18200.00, 85603.50, 'BR-307', 'Robert Chang'],
        ['TXN-2025-00163', '2025-01-27', 'Viktor Ivanov', 7100.00, None, 78503.50, 'BR-101', 'James Liu'],
        ['TXN-2025-00164', '2025-01-28', 'Diana Ramos', 1950.00, None, 76553.50, 'BR-204', 'Priya Sharma'],
        ['TXN-2025-00165', '2025-01-29', 'Chen Wei', None, 11000.00, 87553.50, 'BR-502', 'Nina Petrov'],
        ['TXN-2025-00166', '2025-01-30', 'Emily Watson', 3340.00, None, 84213.50, 'BR-307', 'Robert Chang'],
        ['TXN-2025-00167', '2025-01-31', 'Kwame Asante', 6200.00, None, 78013.50, 'BR-101', 'James Liu'],
        ['TXN-2025-00168', '2025-01-31', 'Julia Fischer', None, 7800.00, 85813.50, 'BR-204', 'Priya Sharma'],
        ['TXN-2025-00169', '2025-01-31', 'Ali Nazari', 2100.00, None, 83713.50, 'BR-502', 'Nina Petrov'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Format currency columns
            if c in (4, 5, 6) and val is not None:
                cell.number_format = '#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
