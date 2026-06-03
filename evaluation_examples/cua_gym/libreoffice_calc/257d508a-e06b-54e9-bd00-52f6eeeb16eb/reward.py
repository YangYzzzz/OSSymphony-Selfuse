"""
Reward Script: Set row 1 height to 1.8cm and enable word wrap for header row
Task ID: calc_gfl_081
Domain: libreoffice_calc
Scoring:
  Component 1: Row 1 height is approximately 1.8cm (51.02 points) — 0.5 points
  Component 2: All 8 cells in row 1 have wrap_text enabled — 0.5 points
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_081'

# 1.8 cm = 1.8 * 28.3465 = 51.0237 points
TARGET_HEIGHT_PT = 51.0237
HEIGHT_TOLERANCE = 2.0  # allow ~2 point tolerance


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Activity Log' sheet must exist
    if 'Activity Log' not in wb.sheetnames:
        print("CRITICAL: 'Activity Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Activity Log']

    # Component 1: Row 1 height is approximately 1.8cm / 51.02 points (0.5 points)
    try:
        row_height = ws.row_dimensions[1].height
        if row_height is not None and abs(row_height - TARGET_HEIGHT_PT) <= HEIGHT_TOLERANCE:
            print(f"PASS: Component 1 — Row 1 height is {row_height:.2f} pts (target ~{TARGET_HEIGHT_PT:.2f} pts = 1.8cm) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Row 1 height is {row_height} (expected ~{TARGET_HEIGHT_PT:.2f} pts = 1.8cm)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 8 cells in row 1 have wrap_text=True (0.5 points)
    # Award partial credit: 0.5 * (cells_with_wrap / 8)
    try:
        wrap_count = 0
        total_header_cells = 8
        for col in range(1, total_header_cells + 1):
            cell = ws.cell(row=1, column=col)
            if cell.alignment.wrap_text is True:
                wrap_count += 1
            else:
                print(f"  DETAIL: Cell row=1 col={col} wrap_text={cell.alignment.wrap_text}")

        if wrap_count == total_header_cells:
            print(f"PASS: Component 2 — All {total_header_cells} header cells have wrap_text=True (0.5 pts)")
            total_score += 0.5
        elif wrap_count > 0:
            partial = 0.5 * (wrap_count / total_header_cells)
            print(f"PARTIAL: Component 2 — {wrap_count}/{total_header_cells} cells have wrap_text=True ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No header cells have wrap_text enabled (0/{total_header_cells})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
