"""
Initial Setup: Chart Y-axis number format task
Task ID: calc_chart_axis_number_format_050
Domain: libreoffice_calc

Creates a spreadsheet with quarterly revenue data (large numbers) and a column chart.
The Y-axis shows full numbers (485000, 523000, etc.) - NOT formatted with K suffix.
"""

import os
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_axis_number_format_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = Workbook()

    # --- Sheet: BigNumbers ---
    ws = wb.active
    ws.title = 'BigNumbers'

    # Headers
    ws['A1'] = 'Quarter'
    ws['B1'] = 'Revenue'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # Data rows: Q1 2024 through Q1 2025 revenue figures
    data = [
        ('Q1 2024', 485000),
        ('Q2 2024', 523000),
        ('Q3 2024', 498000),
        ('Q4 2024', 612000),
        ('Q1 2025', 641000),
    ]
    for r, (quarter, revenue) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=quarter)
        ws.cell(row=r, column=2, value=revenue)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14

    # --- Create a column chart (BarChart type="col") ---
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Quarterly Revenue'
    chart.y_axis.title = 'Revenue ($)'
    chart.x_axis.title = 'Quarter'
    chart.style = 10

    # Data reference (includes header row for series title)
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)

    # Categories (quarter labels)
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.set_categories(cats)

    # Chart dimensions
    chart.width = 18
    chart.height = 12

    # Add chart to sheet (no K-format on Y-axis — that's what the agent must do)
    ws.add_chart(chart, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
