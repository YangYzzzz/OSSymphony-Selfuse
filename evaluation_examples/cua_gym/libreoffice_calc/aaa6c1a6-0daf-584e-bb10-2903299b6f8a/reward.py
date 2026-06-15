"""
Reward Script: Patient Appointment Schedule for Medical Clinic
Task ID: calc_grs_042
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Color coding applied to appointment cells by type
  Component 2 (0.30): Daily summary section with appointment counts by type per doctor
  Component 3 (0.25): Patient Lookup sheet with all appointments listed
  Component 4 (0.15): Data validation (dropdown) on Room Number column
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_042'

# Expected color mapping (ARGB 8-char)
COLOR_MAP = {
    'New Patient': 'FFBDD7EE',    # light blue
    'Follow-up': 'FFC6EFCE',      # light green
    'Procedure': 'FFFCE4D6',      # light orange
    'Emergency': 'FFFF0000',      # red
    'Available': 'FFFFFFFF',      # white
}

# Appointment types to extract from cell text
APPT_TYPES = ['New Patient', 'Follow-up', 'Procedure', 'Emergency']


def get_appt_type(cell_value):
    """Extract appointment type from cell value like 'John Smith - Follow-up'."""
    if cell_value is None:
        return None
    val = str(cell_value).strip()
    if val == 'Available':
        return 'Available'
    for atype in APPT_TYPES:
        if atype in val:
            return atype
    return None


def get_cell_fill_rgb(cell):
    """Get the fill color RGB string of a cell, or None."""
    try:
        if cell.fill.fill_type == 'solid':
            return cell.fill.fgColor.rgb
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that Appointments sheet exists
    if 'Appointments' not in wb.sheetnames:
        print("CRITICAL: 'Appointments' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Appointments']

    # ===================================================================
    # Component 1: Color coding on appointment cells (0.30 points)
    # Check that cells in B4:E39 have correct fill colors based on type
    # This FAILS on initial (no fills) and PASSES on golden (colored fills)
    # ===================================================================
    try:
        correct_colors = 0
        total_cells_with_type = 0

        for row_idx in range(4, 40):  # rows 4-39 (8:00 AM to 4:45 PM)
            for col_idx in range(2, 6):  # columns B-E (4 doctors)
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_val = cell.value
                appt_type = get_appt_type(cell_val)

                if appt_type is not None:
                    total_cells_with_type += 1
                    expected_color = COLOR_MAP.get(appt_type)
                    actual_color = get_cell_fill_rgb(cell)

                    if actual_color == expected_color:
                        correct_colors += 1

        if total_cells_with_type > 0:
            color_ratio = correct_colors / total_cells_with_type
            # Require at least 80% correct for full credit
            if color_ratio >= 0.8:
                component1_score = 0.30
                print(f"PASS: Component 1 - Color coding: {correct_colors}/{total_cells_with_type} cells correct ({color_ratio:.1%}) (0.30 pts)")
            elif color_ratio >= 0.5:
                component1_score = 0.15
                print(f"PARTIAL: Component 1 - Color coding: {correct_colors}/{total_cells_with_type} cells correct ({color_ratio:.1%}) (0.15 pts)")
            else:
                component1_score = 0.0
                print(f"FAIL: Component 1 - Color coding: {correct_colors}/{total_cells_with_type} cells correct ({color_ratio:.1%})")
            total_score += component1_score
        else:
            print("FAIL: Component 1 - No appointment cells found in B4:E39")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # ===================================================================
    # Component 2: Daily summary section with appointment counts (0.30 points)
    # Check for summary rows after the time slots with counts by type per doctor
    # This FAILS on initial (no summary rows) and PASSES on golden (summary present)
    # ===================================================================
    try:
        # Look for a summary section anywhere in rows 40-60
        summary_found = False
        summary_start_row = None

        for row_idx in range(40, 61):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and 'summary' in str(cell_val).lower():
                summary_found = True
                # The header row for types should be 1-2 rows below
                break

        # Look for the type labels in column A after the summary title
        type_rows = {}
        for row_idx in range(40, 61):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val:
                val_str = str(cell_val).strip()
                if val_str in ['New Patient', 'Follow-up', 'Procedure', 'Emergency']:
                    type_rows[val_str] = row_idx

        # Expected counts per doctor from golden data:
        # New Patient: Williams=6 (Margaret2+Angela2+Thomas2), but let's check presence of counts
        # We verify: (a) summary section exists, (b) has rows for all 4 types, (c) has numeric counts
        component2_score = 0.0

        if len(type_rows) >= 4:
            # All four appointment types present in summary
            all_have_counts = True
            for atype, row_idx in type_rows.items():
                # Check B, C, D, E columns have numeric values
                has_number = False
                for col_idx in range(2, 6):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if isinstance(val, (int, float)):
                        has_number = True
                        break
                if not has_number:
                    all_have_counts = False

            if all_have_counts:
                component2_score = 0.30
                print(f"PASS: Component 2 - Daily summary: all 4 types with numeric counts found (0.30 pts)")
            else:
                component2_score = 0.15
                print(f"PARTIAL: Component 2 - Daily summary: types found but some missing counts (0.15 pts)")
        elif len(type_rows) >= 2:
            component2_score = 0.10
            print(f"PARTIAL: Component 2 - Daily summary: only {len(type_rows)}/4 types found (0.10 pts)")
        else:
            print(f"FAIL: Component 2 - Daily summary section not found or incomplete (found {len(type_rows)} type rows)")

        total_score += component2_score
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # ===================================================================
    # Component 3: Patient Lookup sheet (0.25 points)
    # The golden file has a "Patient Lookup" sheet with all appointments listed
    # This FAILS on initial (sheet doesn't exist) and PASSES on golden
    # ===================================================================
    try:
        component3_score = 0.0

        # Check if Patient Lookup sheet exists
        lookup_sheet_exists = False
        lookup_ws = None
        for sn in wb.sheetnames:
            if 'lookup' in sn.lower() or 'patient' in sn.lower():
                if sn != 'Appointments':
                    lookup_sheet_exists = True
                    lookup_ws = wb[sn]
                    break

        if lookup_sheet_exists and lookup_ws is not None:
            # Count data rows (rows with patient names)
            patient_rows = 0
            for row_idx in range(2, lookup_ws.max_row + 1):
                val = lookup_ws.cell(row=row_idx, column=1).value
                if val and str(val).strip() and str(val).strip() not in [
                    'Patient Name', 'Patient Appointment Lookup',
                    'Enter a patient name below to find their appointments',
                    'Patient Name:', 'Color Legend:'
                ]:
                    # Check if it looks like a patient name (has letters, not a header)
                    name = str(val).strip()
                    if any(c.isalpha() for c in name) and name not in ['New Patient', 'Follow-up', 'Procedure', 'Emergency', 'Available']:
                        patient_rows += 1

            if patient_rows >= 30:
                component3_score = 0.25
                print(f"PASS: Component 3 - Patient Lookup sheet: found with {patient_rows} appointment entries (0.25 pts)")
            elif patient_rows >= 10:
                component3_score = 0.15
                print(f"PARTIAL: Component 3 - Patient Lookup sheet: found with {patient_rows} entries, expected 30+ (0.15 pts)")
            elif patient_rows > 0:
                component3_score = 0.05
                print(f"PARTIAL: Component 3 - Patient Lookup sheet: found with only {patient_rows} entries (0.05 pts)")
            else:
                print(f"FAIL: Component 3 - Patient Lookup sheet exists but has no patient data")
        else:
            print(f"FAIL: Component 3 - Patient Lookup sheet not found (sheets: {wb.sheetnames})")

        total_score += component3_score
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # ===================================================================
    # Component 4: Data validation on Room Number column (0.15 points)
    # The task asks for dropdown validation with room options on column F
    # This FAILS on initial (no validations) and PASSES on golden
    # ===================================================================
    try:
        component4_score = 0.0
        has_room_validation = False

        if ws.data_validations and ws.data_validations.dataValidation:
            for dv in ws.data_validations.dataValidation:
                if dv.type == 'list':
                    formula = str(dv.formula1) if dv.formula1 else ''
                    sqref = str(dv.sqref) if dv.sqref else ''
                    # Check if it references column F and has room-related options
                    if 'F' in sqref.upper():
                        # Check formula contains room options
                        formula_lower = formula.lower()
                        if ('exam' in formula_lower or 'procedure' in formula_lower or
                                'consult' in formula_lower or 'room' in formula_lower):
                            has_room_validation = True

        if has_room_validation:
            component4_score = 0.15
            print(f"PASS: Component 4 - Data validation: Room Number dropdown found on column F (0.15 pts)")
        else:
            print(f"FAIL: Component 4 - Data validation: No room dropdown found on column F")

        total_score += component4_score
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
