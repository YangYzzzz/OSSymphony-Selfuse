"""
Initial Setup: Patient appointment schedule for a medical clinic
Task ID: calc_grs_042
Domain: libreoffice_calc

Creates a daily appointment book with time slots from 8am-5pm in 15-min intervals,
4 doctor columns with some appointments filled in. No color-coding, no summary,
no patient lookup sheet, no data validation - those are the task objectives.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_time_slots():
    """Generate 15-minute time slots from 8:00 AM to 5:00 PM."""
    slots = []
    for hour in range(8, 17):  # 8 AM to 4:45 PM
        for minute in [0, 15, 30, 45]:
            h12 = hour if hour <= 12 else hour - 12
            ampm = "AM" if hour < 12 else "PM"
            slots.append(f"{h12}:{minute:02d} {ampm}")
    return slots


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Appointments ---
    ws = wb.active
    ws.title = "Appointments"

    # Title row
    ws.merge_cells("A1:F1")
    ws["A1"] = "Daily Appointment Schedule - Riverside Family Medicine"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Date row
    ws.merge_cells("A2:F2")
    ws["A2"] = "Date: Monday, March 17, 2025"
    ws["A2"].font = Font(name="Calibri", size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers (row 3)
    headers = ["Time Slot", "Dr. Williams", "Dr. Patel", "Dr. Chen", "Dr. Martinez", "Room Number"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Time slots
    time_slots = generate_time_slots()

    # Appointment data - realistic mix of patients and available slots
    # Format: "Patient Name - Type" or "Available"
    appointments = {
        # (row_offset, col): (patient, type, room)
        # Dr. Williams (col 2)
        (0, 2): ("Margaret Thompson - New Patient", "Exam 1"),
        (1, 2): ("Margaret Thompson - New Patient", "Exam 1"),
        (4, 2): ("Robert Garcia - Follow-up", "Exam 1"),
        (6, 2): ("Linda Park - Procedure", "Procedure Room"),
        (7, 2): ("Linda Park - Procedure", "Procedure Room"),
        (8, 2): ("Linda Park - Procedure", "Procedure Room"),
        (10, 2): ("James Wilson - Follow-up", "Exam 2"),
        (14, 2): ("Angela Foster - New Patient", "Exam 1"),
        (15, 2): ("Angela Foster - New Patient", "Exam 1"),
        (18, 2): ("David Kim - Follow-up", "Consult"),
        (22, 2): ("Patricia Moore - Follow-up", "Exam 1"),
        (26, 2): ("Thomas Anderson - New Patient", "Exam 2"),
        (27, 2): ("Thomas Anderson - New Patient", "Exam 2"),
        (30, 2): ("Nancy White - Follow-up", "Exam 1"),
        (34, 2): ("Carlos Rivera - Procedure", "Procedure Room"),
        (35, 2): ("Carlos Rivera - Procedure", "Procedure Room"),

        # Dr. Patel (col 3)
        (0, 3): ("Sarah Chen - Follow-up", "Exam 2"),
        (2, 3): ("Michael Brown - New Patient", "Exam 1"),
        (3, 3): ("Michael Brown - New Patient", "Exam 1"),
        (6, 3): ("Jennifer Liu - Follow-up", "Exam 2"),
        (8, 3): ("William Taylor - Procedure", "Procedure Room"),
        (9, 3): ("William Taylor - Procedure", "Procedure Room"),
        (12, 3): ("Emily Nguyen - Emergency", "Exam 1"),
        (13, 3): ("Emily Nguyen - Emergency", "Exam 1"),
        (16, 3): ("Christopher Hall - Follow-up", "Consult"),
        (20, 3): ("Maria Santos - New Patient", "Exam 2"),
        (21, 3): ("Maria Santos - New Patient", "Exam 2"),
        (24, 3): ("Steven Wright - Follow-up", "Exam 1"),
        (28, 3): ("Amanda Jackson - Procedure", "Procedure Room"),
        (29, 3): ("Amanda Jackson - Procedure", "Procedure Room"),
        (32, 3): ("Daniel Lee - Follow-up", "Exam 2"),
        (35, 3): ("Rachel Green - New Patient", "Consult"),

        # Dr. Chen (col 4)
        (1, 4): ("Barbara Martinez - Procedure", "Procedure Room"),
        (2, 4): ("Barbara Martinez - Procedure", "Procedure Room"),
        (4, 4): ("Kevin O'Brien - Follow-up", "Exam 1"),
        (6, 4): ("Susan Wang - New Patient", "Exam 2"),
        (7, 4): ("Susan Wang - New Patient", "Exam 2"),
        (10, 4): ("Richard Davis - Emergency", "Exam 1"),
        (11, 4): ("Richard Davis - Emergency", "Exam 1"),
        (14, 4): ("Laura Robinson - Follow-up", "Consult"),
        (18, 4): ("Andrew Clark - New Patient", "Exam 1"),
        (19, 4): ("Andrew Clark - New Patient", "Exam 1"),
        (22, 4): ("Michelle Torres - Follow-up", "Exam 2"),
        (24, 4): ("George Miller - Procedure", "Procedure Room"),
        (25, 4): ("George Miller - Procedure", "Procedure Room"),
        (28, 4): ("Diana Hughes - Follow-up", "Exam 1"),
        (32, 4): ("Paul Baker - New Patient", "Consult"),
        (33, 4): ("Paul Baker - New Patient", "Consult"),

        # Dr. Martinez (col 5)
        (0, 5): ("Catherine Phillips - Follow-up", "Exam 1"),
        (2, 5): ("Ronald Scott - Emergency", "Exam 2"),
        (3, 5): ("Ronald Scott - Emergency", "Exam 2"),
        (6, 5): ("Jessica Adams - New Patient", "Exam 1"),
        (7, 5): ("Jessica Adams - New Patient", "Exam 1"),
        (10, 5): ("Mark Evans - Follow-up", "Consult"),
        (12, 5): ("Stephanie Cooper - Procedure", "Procedure Room"),
        (13, 5): ("Stephanie Cooper - Procedure", "Procedure Room"),
        (16, 5): ("Brian Murphy - Follow-up", "Exam 2"),
        (20, 5): ("Donna Reed - New Patient", "Exam 1"),
        (21, 5): ("Donna Reed - New Patient", "Exam 1"),
        (24, 5): ("Frank Russell - Follow-up", "Exam 2"),
        (26, 5): ("Victoria Bell - Procedure", "Procedure Room"),
        (27, 5): ("Victoria Bell - Procedure", "Procedure Room"),
        (30, 5): ("Henry Watson - Emergency", "Exam 1"),
        (31, 5): ("Henry Watson - Emergency", "Exam 1"),
        (34, 5): ("Lisa Morgan - Follow-up", "Consult"),
    }

    # Fill in the schedule
    data_start_row = 4
    cell_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, slot in enumerate(time_slots):
        row = data_start_row + i

        # Time slot
        cell = ws.cell(row=row, column=1, value=slot)
        cell.font = Font(name="Calibri", size=10, bold=True)
        cell.alignment = cell_align
        cell.border = thin_border

        # Doctor columns
        for col in range(2, 6):
            cell = ws.cell(row=row, column=col)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name="Calibri", size=9)

            key = (i, col)
            if key in appointments:
                appt_text = appointments[key][0]
                cell.value = appt_text
            else:
                cell.value = "Available"

        # Room Number column
        room_cell = ws.cell(row=row, column=6)
        room_cell.alignment = cell_align
        room_cell.border = thin_border
        room_cell.font = Font(name="Calibri", size=9)

        # Assign room based on appointments in this row
        rooms_used = []
        for col in range(2, 6):
            key = (i, col)
            if key in appointments:
                rooms_used.append(appointments[key][1])
        if rooms_used:
            room_cell.value = rooms_used[0]  # Show primary room
        else:
            room_cell.value = ""

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 32
    ws.column_dimensions["E"].width = 32
    ws.column_dimensions["F"].width = 16

    # Row heights
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 22

    # Freeze panes - freeze header rows
    ws.freeze_panes = "A4"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
