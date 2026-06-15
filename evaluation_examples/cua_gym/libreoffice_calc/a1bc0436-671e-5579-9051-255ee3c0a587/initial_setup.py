"""
Initial Setup: Enable data table display below chart task
Task ID: calc_chart_data_table_below_062
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.chart import BarChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_data_table_below_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Presentation ---
    ws = wb.active
    ws.title = 'Presentation'

    # Headers in row 1
    headers = ['Product', 'Q1', 'Q2', 'Q3', 'Q4']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows
    data = [
        ['Alpha', 42, 48, 55, 61],
        ['Beta',  38, 41, 45, 52],
        ['Gamma', 29, 34, 39, 43],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Clustered column chart (NO data table) ---
    chart = BarChart()
    chart.type = 'col'          # vertical columns
    chart.grouping = 'clustered'
    chart.title = 'Quarterly Sales by Product'
    chart.y_axis.title = 'Sales Units'
    chart.x_axis.title = 'Quarter'

    # Data reference: columns B-E, rows 1-4 (including header)
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=5, max_row=4)
    cats = Reference(ws, min_col=1, min_row=2, max_row=4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    # No data table in initial state
    # chart.plot_area.dTable is None by default

    chart.width = 20
    chart.height = 14
    ws.add_chart(chart, 'G1')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
