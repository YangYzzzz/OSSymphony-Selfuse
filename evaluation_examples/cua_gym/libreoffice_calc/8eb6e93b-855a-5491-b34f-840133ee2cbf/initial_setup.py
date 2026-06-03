"""
Initial Setup: Monthly report workbook with four sheets (Summary, January, February, March).
Task ID: calc_sht_multiop_003
Domain: libreoffice_calc

Creates a workbook with:
- Summary sheet aggregating monthly data
- January, February, March sheets each with Region + 13 product columns (B-N), rows 2-30
- No tab colors set on any sheet
- No freeze panes on any sheet
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_multiop_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Realistic region names (29 regions to fill rows 2-30)
REGIONS = [
    'North America', 'South America', 'Western Europe', 'Eastern Europe',
    'Nordics', 'Middle East', 'Sub-Saharan Africa', 'North Africa',
    'South Asia', 'Southeast Asia', 'East Asia', 'Australia & NZ',
    'Central Asia', 'Caribbean', 'Central America', 'Iberia',
    'British Isles', 'Benelux', 'DACH', 'France', 'Italy', 'Poland',
    'Russia', 'Turkey', 'Brazil', 'Mexico', 'Canada', 'Japan', 'South Korea'
]

# Realistic product/category column headers (13 products: columns B through N)
PRODUCTS = [
    'Laptops', 'Desktops', 'Tablets', 'Smartphones', 'Monitors',
    'Keyboards', 'Mice', 'Headsets', 'Webcams', 'Printers',
    'Servers', 'Networking', 'Software'
]

# Monthly sales data sets — different base values per month for realism
import random
random.seed(42)

def gen_monthly_data(base_scale):
    """Generate 29 rows x 13 cols of realistic sales figures."""
    data = []
    for i, region in enumerate(REGIONS):
        row = [region]
        for j, prod in enumerate(PRODUCTS):
            # Regional variations: larger regions sell more
            regional_factor = 1.0 + (i % 5) * 0.15
            product_factor = 1.0 + (j % 4) * 0.2
            val = int(base_scale * regional_factor * product_factor * (0.85 + random.random() * 0.3))
            row.append(val)
        data.append(row)
    return data

JAN_DATA = gen_monthly_data(12400)
FEB_DATA = gen_monthly_data(11800)
MAR_DATA = gen_monthly_data(13200)

HEADERS = ['Region'] + PRODUCTS


def create_monthly_sheet(wb, month_name, data):
    """Create a monthly data sheet with realistic content."""
    ws = wb.create_sheet(month_name)

    # Write headers in row 1
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Write data rows 2-30
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    for col_letter in 'BCDEFGHIJKLMN':
        ws.column_dimensions[col_letter].width = 12

    return ws


def create_summary_sheet(wb):
    """Create a Summary sheet that aggregates monthly totals."""
    ws = wb.active
    ws.title = 'Summary'

    # Summary headers
    summary_headers = ['Region', 'Jan Total', 'Feb Total', 'Mar Total', 'Q1 Total']
    for col_idx, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Summary data: region totals across all products
    for row_idx, region in enumerate(REGIONS, 2):
        jan_total = sum(JAN_DATA[row_idx - 2][1:])
        feb_total = sum(FEB_DATA[row_idx - 2][1:])
        mar_total = sum(MAR_DATA[row_idx - 2][1:])
        q1_total = jan_total + feb_total + mar_total

        ws.cell(row=row_idx, column=1, value=region)
        ws.cell(row=row_idx, column=2, value=jan_total)
        ws.cell(row=row_idx, column=3, value=feb_total)
        ws.cell(row=row_idx, column=4, value=mar_total)
        ws.cell(row=row_idx, column=5, value=q1_total)

    # Column widths
    ws.column_dimensions['A'].width = 22
    for col_letter in 'BCDE':
        ws.column_dimensions[col_letter].width = 14

    return ws


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary (active by default) ---
    create_summary_sheet(wb)

    # --- Sheet 2: January ---
    create_monthly_sheet(wb, 'January', JAN_DATA)

    # --- Sheet 3: February ---
    create_monthly_sheet(wb, 'February', FEB_DATA)

    # --- Sheet 4: March ---
    create_monthly_sheet(wb, 'March', MAR_DATA)

    # Verify no tab colors and no freeze panes on any sheet
    for ws in wb.worksheets:
        # No tab color set
        assert ws.sheet_properties.tabColor is None or ws.sheet_properties.tabColor.rgb in (None, '00000000'), \
            f"Sheet {ws.title} has unexpected tab color"
        # No freeze panes
        assert ws.freeze_panes is None, f"Sheet {ws.title} has unexpected freeze panes"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    for ws in wb.worksheets:
        print(f'  {ws.title}: {ws.max_row} rows x {ws.max_column} cols, '
              f'tab_color={ws.sheet_properties.tabColor}, freeze={ws.freeze_panes}')


create_initial()
