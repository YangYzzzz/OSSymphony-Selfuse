"""
Initial Setup: Set default row height and column width in LibreOffice Calc
Task ID: calc_gg1_049
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Data'

    # --- Headers ---
    headers = ['Employee', 'Department', 'Q1 Revenue', 'Q2 Revenue',
               'Start Date', 'Status', 'Region', 'Notes']
    header_font = Font(name='Arial', size=11, bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # --- Realistic data (15 rows) ---
    data = [
        ['Sarah Chen', 'Engineering', 45230.50, 51200.00, '2023-01-15', 'Active', 'West', 'Lead developer on Atlas project'],
        ['Marcus Johnson', 'Marketing', 38750.00, 42100.75, '2022-06-01', 'Active', 'East', 'Manages digital campaigns'],
        ['Priya Patel', 'Finance', 62400.00, 59800.25, '2021-11-20', 'Active', 'Central', 'Senior financial analyst'],
        ['James O\'Brien', 'Sales', 87650.00, 93200.00, '2020-03-10', 'Active', 'West', 'Top performer Q1-Q2'],
        ['Aiko Tanaka', 'Engineering', 52100.00, 54750.50, '2023-04-22', 'Active', 'East', 'Backend infrastructure'],
        ['Carlos Rivera', 'HR', 31500.00, 33200.00, '2022-09-15', 'On Leave', 'Central', 'Parental leave since May'],
        ['Elena Volkov', 'Sales', 71200.00, 68900.00, ' 2021-07-08', 'Active', 'West', 'Enterprise accounts lead'],
        ['David Kim', 'Engineering', 48900.00, 52300.75, '2023-08-01', 'Active', 'East', 'Frontend specialist'],
        ['Fatima Al-Hassan', 'Finance', 55800.00, 58100.00, '2022-02-14', 'Active', 'Central', 'Budget planning coordinator'],
        ['Ryan Mitchell', 'Marketing', 42300.00, 44800.50, '2021-12-03', 'Active', 'West', 'Content strategy manager'],
        ['Mei Lin', 'Sales', 63400.00, 71500.00, '2020-10-19', 'Active', 'East', 'APAC regional sales'],
        ['Thomas Bergmann', 'Engineering', 57200.00, 59100.25, '2022-05-30', 'Active', 'Central', 'DevOps and CI/CD'],
        ['Olivia Foster', 'HR', 34800.00, 36100.00, '2023-03-07', 'Active', 'West', 'Recruitment specialist'],
        ['Raj Gupta', 'Finance', 49500.00, 51700.50, '2021-09-12', 'Active', 'East', 'Tax compliance officer'],
        ['Sophie Laurent', 'Marketing', 41200.00, 45600.75, '2022-11-25', 'Active', 'Central', 'Brand partnerships lead'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Set INCONSISTENT row heights (simulating pasted data from multiple sources) ---
    ws.row_dimensions[1].height = 25     # header row too tall
    ws.row_dimensions[2].height = 35     # leftover from large font paste
    ws.row_dimensions[3].height = 15     # default-ish
    ws.row_dimensions[4].height = 40     # very tall from external paste
    ws.row_dimensions[5].height = 12     # too short
    ws.row_dimensions[6].height = 30     # tall
    ws.row_dimensions[7].height = 18     # slightly short
    ws.row_dimensions[8].height = 38     # tall
    ws.row_dimensions[9].height = 14     # short
    ws.row_dimensions[10].height = 32    # tall
    ws.row_dimensions[11].height = 16    # short
    ws.row_dimensions[12].height = 42    # very tall
    ws.row_dimensions[13].height = 13    # short
    ws.row_dimensions[14].height = 28    # medium tall
    ws.row_dimensions[15].height = 36    # tall
    ws.row_dimensions[16].height = 22    # slightly tall

    # --- Set INCONSISTENT column widths (text cut off in some, too wide in others) ---
    ws.column_dimensions['A'].width = 10   # too narrow for names
    ws.column_dimensions['B'].width = 8    # too narrow, text cut off
    ws.column_dimensions['C'].width = 20   # wider than needed
    ws.column_dimensions['D'].width = 9    # too narrow
    ws.column_dimensions['E'].width = 25   # overly wide
    ws.column_dimensions['F'].width = 7    # very narrow
    ws.column_dimensions['G'].width = 12   # slightly narrow
    ws.column_dimensions['H'].width = 30   # way too wide for notes

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
