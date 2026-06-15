"""
Reward Script: Set default row height to 20pt and column width A-H to 15 characters
Task ID: calc_gg1_049
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): All rows in 'Data' sheet have height == 20.0 pt
  Component 2 (0.35): Columns A-H in 'Data' sheet have width == 15.0 characters
  Component 3 (0.15): Cell data integrity preserved (spot-check key cells)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_049'


def persist_app_state(domain: str):
    """Attempt to save any unsaved GUI state via Ctrl+S."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Data' sheet must exist
    if 'Data' not in wb.sheetnames:
        print(f"CRITICAL: 'Data' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data']
    max_row = ws.max_row
    if max_row is None or max_row < 1:
        print("CRITICAL: Sheet has no data rows")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All rows have height == 20.0 (0.50 points)
    # This checks the task-introduced change: row heights were inconsistent, now must be 20pt
    try:
        correct_rows = 0
        total_rows = max_row
        tolerance = 0.5  # allow tiny floating point variance
        for r in range(1, total_rows + 1):
            h = ws.row_dimensions[r].height
            if h is not None and abs(h - 20.0) <= tolerance:
                correct_rows += 1
            else:
                print(f"  Row {r}: height={h} (expected ~20.0)")

        if correct_rows == total_rows:
            print(f"PASS: Component 1 — All {total_rows} rows have height 20.0 pt (0.50 pts)")
            total_score += 0.50
        elif correct_rows > 0:
            # Partial credit: proportional to rows correct
            fraction = correct_rows / total_rows
            partial = round(0.50 * fraction, 3)
            print(f"PARTIAL: Component 1 — {correct_rows}/{total_rows} rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No rows have height 20.0. 0/{total_rows} correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Columns A-H have width == 15.0 (0.35 points)
    # This checks the task-introduced change: column widths were inconsistent, now must be 15 characters
    try:
        target_cols = list("ABCDEFGH")
        correct_cols = 0
        col_tolerance = 0.5
        for col in target_cols:
            w = ws.column_dimensions[col].width
            if w is not None and abs(w - 15.0) <= col_tolerance:
                correct_cols += 1
            else:
                print(f"  Col {col}: width={w} (expected ~15.0)")

        if correct_cols == len(target_cols):
            print(f"PASS: Component 2 — All columns A-H have width 15.0 chars (0.35 pts)")
            total_score += 0.35
        elif correct_cols > 0:
            fraction = correct_cols / len(target_cols)
            partial = round(0.35 * fraction, 3)
            print(f"PARTIAL: Component 2 — {correct_cols}/{len(target_cols)} columns correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No columns have width 15.0. 0/{len(target_cols)} correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data integrity preserved (0.15 points)
    # The task says cell content should be preserved. Verify key cells still have expected values.
    # These values are only scored IF row heights AND column widths changed (i.e., task was attempted).
    # On initial_env, rows/cols are wrong so this component should not fully pass.
    # We gate this: only award if at least one of Component 1 or 2 earned partial credit.
    try:
        # Check if any task change was made (gate condition)
        if total_score > 0:
            expected_cells = {
                'A1': 'Employee',
                'B1': 'Department',
                'C1': 'Q1 Revenue',
                'A2': 'Sarah Chen',
                'B2': 'Engineering',
            }
            intact_count = 0
            for coord, expected_val in expected_cells.items():
                actual = ws[coord].value
                if actual is not None and str(actual).strip() == expected_val:
                    intact_count += 1
                else:
                    print(f"  Cell {coord}: expected '{expected_val}', found '{actual}'")

            if intact_count == len(expected_cells):
                print(f"PASS: Component 3 — Data integrity preserved, {intact_count}/{len(expected_cells)} cells intact (0.15 pts)")
                total_score += 0.15
            else:
                fraction = intact_count / len(expected_cells)
                partial = round(0.15 * fraction, 3)
                print(f"PARTIAL: Component 3 — {intact_count}/{len(expected_cells)} cells intact ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 3 — Skipped (no task changes detected, gate condition not met)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
