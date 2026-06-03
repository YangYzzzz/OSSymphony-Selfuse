"""
Reward Script: Sheet protection on 'Config' with selective cell locking
Task ID: calc_ps_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Sheet protection is enabled on 'Config'
  Component 2 (0.2): selectLockedCells is True (users cannot select locked cells)
  Component 3 (0.3): Cell F2 is the only unlocked cell
  Component 4 (0.2): Protection password is set
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_033'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Config' sheet must exist
    if 'Config' not in wb.sheetnames:
        print(f"FAIL: 'Config' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Config']

    # Component 1: Sheet protection is enabled (0.3 points)
    # Initial: sheet=False -> FAIL. Golden: sheet=True -> PASS.
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 -- Sheet protection is enabled (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 -- Sheet protection is not enabled (protection.sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: selectLockedCells is True (0.2 points)
    # In openpyxl, selectLockedCells=True means users CANNOT select locked cells.
    # Initial: selectLockedCells=False -> FAIL. Golden: selectLockedCells=True -> PASS.
    try:
        if ws.protection.selectLockedCells:
            print(f"PASS: Component 2 -- selectLockedCells is True (users cannot select locked cells) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 -- selectLockedCells is False (users can still select locked cells)")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Cell F2 is the only unlocked cell in the used range (0.3 points)
    # Initial: F2 locked=True (all cells locked) -> FAIL. Golden: F2 locked=False, all others locked -> PASS.
    try:
        f2_unlocked = not ws['F2'].protection.locked
        # Check that no other cells in A1:F20 are unlocked
        other_unlocked = []
        for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=6):
            for cell in row:
                if cell.coordinate != 'F2' and not cell.protection.locked:
                    other_unlocked.append(cell.coordinate)

        if f2_unlocked and len(other_unlocked) == 0:
            print(f"PASS: Component 3 -- F2 is unlocked and all other cells are locked (0.3 pts)")
            total_score += 0.3
        elif f2_unlocked and len(other_unlocked) > 0:
            # Partial: F2 is correct but other cells also unlocked
            print(f"PARTIAL: Component 3 -- F2 is unlocked but other cells also unlocked: {other_unlocked[:5]} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 -- F2 locked={ws['F2'].protection.locked}, other unlocked={other_unlocked[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Protection password is set (0.2 points)
    # Initial: password=None -> FAIL. Golden: password is set (hashed) -> PASS.
    try:
        pwd = ws.protection.password
        if pwd is not None and pwd != '':
            print(f"PASS: Component 4 -- Protection password is set (hash: {pwd}) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 -- No protection password set (password={pwd})")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
