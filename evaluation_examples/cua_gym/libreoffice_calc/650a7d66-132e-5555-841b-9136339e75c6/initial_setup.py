"""
Initial Setup: Sheet protection task - Config sheet with configuration data
Task ID: calc_ps_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Config'

    # --- Styling definitions ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    section_font = Font(name="Calibri", size=11, bold=True, color="2F5496")
    section_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    value_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 20

    # --- Row 1: Headers ---
    headers = ['Parameter', 'Value', 'Parameter', 'Value', 'Parameter', 'Value']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # --- Configuration data organized in 3 column pairs ---
    # Column A-B: Database Settings
    # Column C-D: Network Settings
    # Column E-F: Application Settings

    config_data = [
        # Row 2
        ['DB Host', 'db-prod-01.internal', 'Gateway IP', '10.0.1.1', 'Max Connections', 100],
        # Row 3
        ['DB Port', 5432, 'Subnet Mask', '255.255.255.0', 'Session Timeout (s)', 3600],
        # Row 4
        ['DB Name', 'appdata_prod', 'DNS Primary', '8.8.8.8', 'Log Level', 'INFO'],
        # Row 5
        ['Max Pool Size', 25, 'DNS Secondary', '8.8.4.4', 'Cache TTL (min)', 15],
        # Row 6
        ['Connection Timeout', 30, 'Proxy Host', 'proxy.corp.net', 'Thread Pool Size', 8],
        # Row 7
        ['Read Replica', 'db-read-01.internal', 'Proxy Port', 8080, 'Retry Count', 3],
        # Row 8
        ['SSL Enabled', 'TRUE', 'VPN Endpoint', 'vpn.corp.net', 'Batch Size', 500],
        # Row 9
        ['Backup Schedule', 'Daily 02:00 UTC', 'Firewall Zone', 'DMZ', 'Rate Limit (req/s)', 1000],
        # Row 10
        ['Retention Days', 90, 'Load Balancer', 'lb-prod-01', 'API Version', 'v2.4.1'],
        # Row 11
        ['Replication Mode', 'Async', 'CDN Provider', 'CloudFront', 'Feature Flags', 'Enabled'],
        # Row 12
        ['Charset', 'UTF-8', 'TLS Version', '1.3', 'Debug Mode', 'FALSE'],
        # Row 13
        ['Auto Vacuum', 'ON', 'CORS Origin', '*.corp.net', 'Maintenance Window', 'Sun 03:00'],
        # Row 14
        ['WAL Level', 'replica', 'HTTP Timeout (s)', 30, 'Telemetry', 'Enabled'],
        # Row 15
        ['Temp Tablespace', 'pg_default', 'Max Upload (MB)', 50, 'Locale', 'en_US.UTF-8'],
        # Row 16
        ['Shared Buffers', '4GB', 'WebSocket Port', 8443, 'Timezone', 'UTC'],
        # Row 17
        ['Work Mem', '256MB', 'SMTP Host', 'smtp.corp.net', 'Notification Queue', 'SQS'],
        # Row 18
        ['Effective Cache', '12GB', 'SMTP Port', 587, 'Worker Processes', 4],
        # Row 19
        ['Checkpoint Interval', '5 min', 'SMTP Auth', 'STARTTLS', 'Healthcheck Path', '/api/health'],
        # Row 20
        ['Stats Collector', 'ON', 'Alert Email', 'ops@corp.net', 'Deployment Region', 'us-east-1'],
    ]

    for r, row_data in enumerate(config_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = value_font
            cell.border = thin_border
            # Parameter columns (A, C, E) get section styling
            if c in (1, 3, 5):
                cell.font = section_font
                cell.fill = section_fill

    # All cells are locked by default in openpyxl (Protection(locked=True))
    # Sheet is NOT protected - this is the default state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
