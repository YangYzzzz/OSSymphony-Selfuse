"""
Initial Setup: Line chart with no axis titles (agent task: add axis titles)
Task ID: calc_chart_axis_titles_018
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.chart import LineChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_axis_titles_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Climate ---
    ws = wb.active
    ws.title = 'Climate'

    # Headers
    ws['A1'] = 'Month'
    ws['B1'] = 'Temperature'

    # Monthly average temperature data (Jan - Dec)
    monthly_data = [
        ('Jan', 4.2),
        ('Feb', 5.8),
        ('Mar', 9.1),
        ('Apr', 14.3),
        ('May', 18.7),
        ('Jun', 22.4),
        ('Jul', 25.1),
        ('Aug', 24.6),
        ('Sep', 20.2),
        ('Oct', 14.8),
        ('Nov', 9.3),
        ('Dec', 5.1),
    ]

    for row_idx, (month, temp) in enumerate(monthly_data, 2):
        ws.cell(row=row_idx, column=1, value=month)
        ws.cell(row=row_idx, column=2, value=temp)

    # --- Create Line Chart (no axis titles — that's the task) ---
    chart = LineChart()
    chart.title = 'Monthly Average Temperature'
    chart.style = 10
    # Do NOT set axis titles — the agent task is to add them

    # Data reference: B1:B13 (includes header for series name)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=13)
    chart.add_data(data_ref, titles_from_data=True)

    # Category reference: A2:A13 (month names)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=13)
    chart.set_categories(cats_ref)

    # Position chart below data
    ws.add_chart(chart, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Climate sheet: Month/Temperature data for Jan-Dec')
    print('Line chart: "Monthly Average Temperature" — NO axis titles (to be added by agent)')


create_initial()
