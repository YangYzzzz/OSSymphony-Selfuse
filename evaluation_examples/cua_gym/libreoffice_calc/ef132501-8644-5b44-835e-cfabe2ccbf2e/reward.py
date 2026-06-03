"""
Reward Script: Calculate variance (Actual - Budget) for each department and create Sheet2 summary
Task ID: osworld_calc_gross_profit_sheet2_concat_012
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Column D (D2:D15) has variance formulas =C#-B# for all 14 data rows  (0.6 pts)
  Component 2: Sheet2 A1 has a formula containing 'Total Variance' text and
               references SUM(Sheet1!D2:D15)                                          (0.4 pts)
  Total: 1.0
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_012'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 must exist
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']

    # -----------------------------------------------------------------------
    # Component 1: Column D (D2:D15) has variance formulas =C#-B# (0.6 pts)
    #
    # In the initial env, D2:D15 are all empty (None).
    # In the golden env, every cell D2:D15 contains a formula like =C2-B2.
    # We require at least all 14 rows to have a variance formula.
    # -----------------------------------------------------------------------
    try:
        expected_rows = list(range(2, 16))  # rows 2..15 inclusive (14 rows)
        formula_pattern = re.compile(r'^=\s*C(\d+)\s*-\s*B(\d+)\s*$', re.IGNORECASE)

        rows_with_correct_formula = 0
        rows_checked = 0

        for r in expected_rows:
            cell = ws1.cell(row=r, column=4)  # Column D
            val = cell.value
            rows_checked += 1
            if val is None:
                print(f"FAIL: D{r} is empty — expected variance formula =C{r}-B{r}")
                continue
            if not isinstance(val, str):
                print(f"FAIL: D{r} is not a formula string (got {val!r})")
                continue
            m = formula_pattern.match(val.strip())
            if m:
                row_num_c = int(m.group(1))
                row_num_b = int(m.group(2))
                if row_num_c == r and row_num_b == r:
                    rows_with_correct_formula += 1
                else:
                    print(f"FAIL: D{r} formula references wrong rows: {val!r}")
            else:
                # Accept equivalent formulas that still compute Actual - Budget
                # e.g. =(C2-B2) or =C2-B2 with spaces
                stripped = val.strip().lstrip('=').strip()
                alt_pattern = re.compile(
                    r'^\(?\s*C(\d+)\s*-\s*B(\d+)\s*\)?$', re.IGNORECASE
                )
                m2 = alt_pattern.match(stripped)
                if m2 and int(m2.group(1)) == r and int(m2.group(2)) == r:
                    rows_with_correct_formula += 1
                else:
                    print(f"FAIL: D{r} has unexpected formula: {val!r}")

        if rows_with_correct_formula == len(expected_rows):
            print(
                f"PASS: Component 1 — all {rows_with_correct_formula}/14 variance formulas "
                f"present in D2:D15 (0.6 pts)"
            )
            total_score += 0.6
        elif rows_with_correct_formula > 0:
            # Partial: give proportional credit within Component 1
            partial = round(0.6 * rows_with_correct_formula / len(expected_rows), 4)
            print(
                f"PARTIAL: Component 1 — {rows_with_correct_formula}/14 variance formulas "
                f"present in D2:D15 ({partial} pts)"
            )
            total_score += partial
        else:
            print("FAIL: Component 1 — no variance formulas found in D2:D15 (0.0 pts)")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Sheet2 A1 contains a formula with "Total Variance" and
    #               SUM(Sheet1!D2:D15)  (0.4 pts)
    #
    # In the initial env, Sheet2 A1 is empty.
    # In the golden env, Sheet2 A1 is:
    #   ="Total Variance: $"&TEXT(SUM(Sheet1!D2:D15),"#,##0.00")
    # We require:
    #   (a) The cell starts with '=' (it's a formula)
    #   (b) It contains the literal text "Total Variance" (case-insensitive)
    #   (c) It references SUM(Sheet1!D2:D15) (covers D-column range on Sheet1)
    # -----------------------------------------------------------------------
    try:
        if 'Sheet2' not in wb.sheetnames:
            print("FAIL: Component 2 — Sheet2 does not exist")
        else:
            ws2 = wb['Sheet2']
            a1_val = ws2['A1'].value

            if a1_val is None:
                print("FAIL: Component 2 — Sheet2 A1 is empty (0.0 pts)")
            elif not isinstance(a1_val, str) or not a1_val.strip().startswith('='):
                print(
                    f"FAIL: Component 2 — Sheet2 A1 is not a formula (got {a1_val!r}) (0.0 pts)"
                )
            else:
                formula_upper = a1_val.upper().replace(' ', '')
                has_total_variance = 'TOTALVARIANCE' in formula_upper
                # Accept any SUM over Sheet1!D column range
                has_sum_ref = bool(
                    re.search(r'SUM\s*\(\s*Sheet1\s*!\s*D\d+\s*:\s*D\d+\s*\)',
                              a1_val, re.IGNORECASE)
                )
                # Also accept without quoting the sheet name (Sheet1!D2:D15)
                if not has_sum_ref:
                    has_sum_ref = bool(
                        re.search(r'SUM\s*\(.*D\d+\s*:\s*D\d+.*\)',
                                  a1_val, re.IGNORECASE)
                    )

                if has_total_variance and has_sum_ref:
                    print(
                        f"PASS: Component 2 — Sheet2 A1 has 'Total Variance' formula "
                        f"with SUM(Sheet1!D...) reference (0.4 pts)"
                    )
                    total_score += 0.4
                elif has_total_variance:
                    print(
                        f"PARTIAL: Component 2 — Sheet2 A1 has 'Total Variance' text "
                        f"but missing correct SUM(Sheet1!D2:D15) reference; "
                        f"formula: {a1_val!r} (0.2 pts)"
                    )
                    total_score += 0.2
                elif has_sum_ref:
                    print(
                        f"PARTIAL: Component 2 — Sheet2 A1 has SUM(Sheet1!D...) reference "
                        f"but missing 'Total Variance' text; "
                        f"formula: {a1_val!r} (0.2 pts)"
                    )
                    total_score += 0.2
                else:
                    print(
                        f"FAIL: Component 2 — Sheet2 A1 formula does not meet requirements: "
                        f"{a1_val!r} (0.0 pts)"
                    )

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
