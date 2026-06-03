"""
Initial Setup: Department Budget vs Actual table with empty Variance column
Task ID: osworld_calc_gross_profit_sheet2_concat_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: BudgetActual ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers: Department, Budget, Actual, Variance (D is empty — task is to fill it)
    headers = ['Department', 'Budget', 'Actual', 'Variance']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic department budget data (14 rows, rows 2–15)
    # Variance column D intentionally left empty
    data = [
        ['Engineering',        485000, 512300],
        ['Marketing',          210000, 198750],
        ['Sales',              320000, 347200],
        ['Human Resources',     95000,  88600],
        ['Finance',            145000, 151400],
        ['Operations',         275000, 268900],
        ['Customer Support',   130000, 141800],
        ['Research & Dev',     390000, 402100],
        ['Legal',               72000,  69500],
        ['IT Infrastructure',  180000, 175300],
        ['Product Management',  98000, 104700],
        ['Supply Chain',       165000, 158400],
        ['Quality Assurance',   87000,  91200],
        ['Executive Office',   220000, 215600],
    ]

    for r, row_data in enumerate(data, 2):
        ws1.cell(row=r, column=1, value=row_data[0])  # Department
        ws1.cell(row=r, column=2, value=row_data[1])  # Budget
        ws1.cell(row=r, column=3, value=row_data[2])  # Actual
        # Column D (Variance) intentionally left EMPTY — task is to fill it

    # --- Sheet 2: Summary (initially empty, task will add total variance formula) ---
    ws2 = wb.create_sheet('Sheet2')
    # A1 intentionally left EMPTY — task is to add "Total Variance: $X" formula

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
