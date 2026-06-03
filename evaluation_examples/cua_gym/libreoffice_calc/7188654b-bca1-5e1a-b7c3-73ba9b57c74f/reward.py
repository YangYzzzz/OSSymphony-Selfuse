"""
Reward Script: CrossRef lookup formula verification
Task ID: calc_mcp_057
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4 pts): C2 contains a formula (starts with '=')
  Component 2 (0.3 pts): Formula uses INDEX+MATCH pattern referencing Catalog sheet
  Component 3 (0.3 pts): Formula correctly references Catalog.E for return and Catalog.A for lookup
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_057'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: CrossRef sheet must exist
    if 'CrossRef' not in wb.sheetnames:
        print("FAIL: 'CrossRef' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CrossRef']

    # Get the value of C2
    c2_value = ws['C2'].value
    print(f"DEBUG: CrossRef C2 raw value = {c2_value!r}")

    # Component 1: C2 contains a formula (0.4 points)
    # This FAILS on initial (C2 is None) and PASSES on golden (C2 has formula)
    try:
        if c2_value is not None and isinstance(c2_value, str) and c2_value.startswith('='):
            print(f"PASS: Component 1 -- C2 contains a formula: {c2_value} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- C2 does not contain a formula, found: {c2_value!r}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula uses INDEX+MATCH pattern referencing Catalog (0.3 points)
    # This FAILS on initial (no formula) and PASSES on golden (INDEX+MATCH with Catalog ref)
    try:
        if c2_value is not None and isinstance(c2_value, str):
            formula_upper = c2_value.upper().replace(" ", "")
            has_index = 'INDEX(' in formula_upper
            has_match = 'MATCH(' in formula_upper
            # Accept both "Catalog." (LibreOffice) and "Catalog!" (Excel) sheet reference syntax
            has_catalog_ref = 'CATALOG.' in formula_upper or 'CATALOG!' in formula_upper
            if has_index and has_match and has_catalog_ref:
                print(f"PASS: Component 2 -- Formula uses INDEX+MATCH with Catalog reference (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not has_index:
                    missing.append('INDEX')
                if not has_match:
                    missing.append('MATCH')
                if not has_catalog_ref:
                    missing.append('Catalog sheet reference')
                print(f"FAIL: Component 2 -- Formula missing: {', '.join(missing)}")
        else:
            print(f"FAIL: Component 2 -- C2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Formula references correct columns - Catalog E for return values, Catalog A for lookup (0.3 points)
    # This FAILS on initial (no formula) and PASSES on golden
    try:
        if c2_value is not None and isinstance(c2_value, str):
            formula_upper = c2_value.upper().replace(" ", "")
            # Check for Catalog.E (or Catalog!E) column reference for INDEX return array
            has_col_e = bool(re.search(r'CATALOG[.!]E', formula_upper))
            # Check for Catalog.A (or Catalog!A) column reference for MATCH lookup array
            has_col_a = bool(re.search(r'CATALOG[.!]A', formula_upper))
            if has_col_e and has_col_a:
                print(f"PASS: Component 3 -- Formula references Catalog.E (return) and Catalog.A (lookup) (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not has_col_e:
                    missing.append('Catalog column E reference')
                if not has_col_a:
                    missing.append('Catalog column A reference')
                print(f"FAIL: Component 3 -- Formula missing: {', '.join(missing)}")
        else:
            print(f"FAIL: Component 3 -- C2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
