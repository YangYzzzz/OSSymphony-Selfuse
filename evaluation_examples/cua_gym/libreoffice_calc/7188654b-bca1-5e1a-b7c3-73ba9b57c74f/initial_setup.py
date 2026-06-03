"""
Initial Setup: Create a spreadsheet with CrossRef and Catalog sheets for INDEX/MATCH lookup task.
Task ID: calc_mcp_057
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Catalog (product database) ---
    ws_cat = wb.active
    ws_cat.title = 'Catalog'

    headers_cat = ['Product', 'Category', 'SKU', 'Description', 'Cost']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers_cat, 1):
        cell = ws_cat.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    catalog_data = [
        ['Bolt Assembly M8', 'Hardware', 'HW-1001', 'M8 hex bolt with nut and washer', 3.45],
        ['Copper Wire 14AWG', 'Electrical', 'EL-2034', '14 AWG solid copper wire per meter', 1.82],
        ['Widget X', 'Components', 'CP-3078', 'Standard precision widget assembly', 24.95],
        ['Thermal Paste TG-7', 'Thermal', 'TH-4501', 'High conductivity thermal compound 5g', 8.50],
        ['LED Panel 600x600', 'Lighting', 'LT-5200', '40W flat panel ceiling light 4000K', 45.00],
        ['Nylon Spacer M4', 'Hardware', 'HW-1055', 'M4 nylon standoff spacer 10mm', 0.35],
        ['PVC Conduit 25mm', 'Electrical', 'EL-2089', '25mm rigid PVC conduit per meter', 2.10],
        ['Silicone Gasket Ring', 'Seals', 'SL-6010', 'High-temp silicone O-ring 50mm OD', 1.95],
        ['Stainless Hinge 75mm', 'Hardware', 'HW-1102', '75mm butt hinge stainless steel', 6.20],
        ['Carbon Filter CF-12', 'Filtration', 'FT-7003', 'Activated carbon filter 12-inch', 18.75],
        ['Aluminium Bracket L', 'Hardware', 'HW-1200', 'L-shaped mounting bracket 100mm', 4.60],
        ['Fiber Optic Patch 3m', 'Networking', 'NT-8010', 'SC-SC single mode patch cable 3m', 12.30],
    ]

    for r, row_data in enumerate(catalog_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_cat.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws_cat.column_dimensions['A'].width = 22
    ws_cat.column_dimensions['B'].width = 14
    ws_cat.column_dimensions['C'].width = 12
    ws_cat.column_dimensions['D'].width = 40
    ws_cat.column_dimensions['E'].width = 10

    # Format cost column as currency
    for r in range(2, 2 + len(catalog_data)):
        ws_cat.cell(row=r, column=5).number_format = '$#,##0.00'

    # --- Sheet 2: CrossRef (lookup sheet) ---
    ws_cr = wb.create_sheet('CrossRef')

    headers_cr = ['Product', 'Qty', 'Unit Cost']
    for col, h in enumerate(headers_cr, 1):
        cell = ws_cr.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # A2 = Widget X, B2 = quantity, C2 = EMPTY (this is where the formula goes)
    ws_cr.cell(row=2, column=1, value='Widget X')
    ws_cr.cell(row=2, column=2, value=150)
    # C2 intentionally left empty - the task is to add the lookup formula here

    # Add a few more rows to make it realistic
    ws_cr.cell(row=3, column=1, value='Bolt Assembly M8')
    ws_cr.cell(row=3, column=2, value=500)
    # C3 also empty

    ws_cr.cell(row=4, column=1, value='LED Panel 600x600')
    ws_cr.cell(row=4, column=2, value=20)
    # C4 also empty

    ws_cr.cell(row=5, column=1, value='Silicone Gasket Ring')
    ws_cr.cell(row=5, column=2, value=1000)
    # C5 also empty

    ws_cr.column_dimensions['A'].width = 22
    ws_cr.column_dimensions['B'].width = 10
    ws_cr.column_dimensions['C'].width = 12

    # Format Qty as integer and Unit Cost as currency
    for r in range(2, 6):
        ws_cr.cell(row=r, column=2).number_format = '0'
        ws_cr.cell(row=r, column=3).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
