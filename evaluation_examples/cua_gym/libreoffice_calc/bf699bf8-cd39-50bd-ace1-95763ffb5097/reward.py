"""
Reward Script: Calculate bi-weekly payment for a personal loan using PMT formula
Task ID: calc_fmb_pmt_biweekly_058
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: B6 contains a PMT formula (0.5 pts)
  Component 2: PMT formula correctly references B3/B5 for rate, B4*B5 for nper, -B2 for pv (0.5 pts)
  Total: 1.0

The task requires inserting =PMT(B3/B5,B4*B5,-B2) into cell B6 of the 'Personal Loan' sheet.
Initial state: B6 is empty.
Golden state: B6 contains =PMT(B3/B5,B4*B5,-B2).
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_pmt_biweekly_058'


def normalize_formula(formula):
    """Remove whitespace and convert to uppercase for comparison."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Check that 'Personal Loan' sheet exists
    if 'Personal Loan' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Personal Loan' not found. Sheets present: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Personal Loan']

    # Component 1: B6 contains a PMT formula (0.5 points)
    # The key task change is that B6 transitions from empty to containing a PMT formula.
    # This check FAILS on the initial file (B6 is empty) and PASSES on the golden file.
    try:
        b6_value = ws['B6'].value
        if b6_value is None or not isinstance(b6_value, str):
            print(f"FAIL: Component 1 — B6 is empty or not a formula string; found: {repr(b6_value)}")
        else:
            normalized = normalize_formula(b6_value)
            if normalized.startswith('=PMT('):
                print(f"PASS: Component 1 — B6 contains a PMT formula: {repr(b6_value)} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 — B6 does not start with =PMT(; found: {repr(b6_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: PMT formula correctly encodes bi-weekly payment logic (0.5 points)
    # The formula must use:
    #   - B3/B5 as the periodic interest rate (annual rate / payments per year)
    #   - B4*B5 as the total number of payments (years * payments per year)
    #   - -B2 (or negative sign on loan amount) as present value
    # Accepted canonical formula: =PMT(B3/B5,B4*B5,-B2)
    # This check also FAILS on the initial file (B6 is empty) and PASSES on the golden file.
    try:
        b6_value = ws['B6'].value
        if b6_value is None or not isinstance(b6_value, str):
            print(f"FAIL: Component 2 — B6 is empty or not a string; cannot verify formula correctness")
        else:
            normalized = normalize_formula(b6_value)
            # Check that the formula contains the three expected arguments in the right structure:
            # rate  = B3/B5
            # nper  = B4*B5
            # pv    = -B2
            has_rate = 'B3/B5' in normalized
            has_nper = 'B4*B5' in normalized
            # Accept -B2 either as the pv argument or as a negated reference
            has_pv = '-B2' in normalized or 'B2*-1' in normalized

            if has_rate and has_nper and has_pv:
                print(f"PASS: Component 2 — PMT formula correctly encodes B3/B5 (rate), B4*B5 (nper), -B2 (pv): {repr(b6_value)} (0.5 pts)")
                total_score += 0.5
            else:
                missing = []
                if not has_rate:
                    missing.append('B3/B5 (rate = annual_rate/payments_per_year)')
                if not has_nper:
                    missing.append('B4*B5 (nper = years*payments_per_year)')
                if not has_pv:
                    missing.append('-B2 (pv = negative loan amount)')
                print(f"FAIL: Component 2 — Formula missing correct arguments: {missing}. Found: {repr(b6_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
