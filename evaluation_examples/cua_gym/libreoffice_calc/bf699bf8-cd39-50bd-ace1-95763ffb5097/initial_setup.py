"""
Initial Setup: Calculate bi-weekly payment for a personal loan using PMT
Task ID: calc_fmb_pmt_biweekly_058
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'calc_fmb_pmt_biweekly_058'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Personal Loan ---
    ws = wb.active
    ws.title = 'Personal Loan'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18

    # Row 1: Title header
    ws['A1'] = 'Personal Loan Calculator'
    ws['A1'].font = Font(bold=True, size=14)

    # Row 2: Loan Amount
    ws['A2'] = 'Loan Amount'
    ws['B2'] = 15000

    # Row 3: Annual Rate
    ws['A3'] = 'Annual Rate'
    ws['B3'] = 0.085

    # Row 4: Years
    ws['A4'] = 'Years'
    ws['B4'] = 3

    # Row 5: Payments per Year
    ws['A5'] = 'Payments per Year'
    ws['B5'] = 26

    # Row 6: Bi-Weekly Payment (target cell — B6 is intentionally empty)
    ws['A6'] = 'Bi-Weekly Payment'
    # B6 is empty — the agent must add the PMT formula here

    # Number formatting
    ws['B2'].number_format = '$#,##0.00'
    ws['B3'].number_format = '0.00%'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
