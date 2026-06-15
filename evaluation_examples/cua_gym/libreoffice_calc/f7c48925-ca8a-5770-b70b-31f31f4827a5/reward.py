"""
Reward Script: Create a column chart from the monthly marketing spend table
Task ID: osworld_calc_dual_chart_separate_tables_009
Domain: libreoffice_calc
Scoring:
  Component 1: A chart exists in the 'Marketing Data' sheet (0.4 pts)
  Component 2: The chart is a column chart (BarChart with type='col') (0.3 pts)
  Component 3: The chart data references the marketing spend table (rows 2-9, col B) (0.3 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_dual_chart_separate_tables_009'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Create a column chart from the monthly marketing spend table (rows 1-9).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook — if this fails, score is 0
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: the expected sheet must exist
    target_sheet = 'Marketing Data'
    if target_sheet not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{target_sheet}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[target_sheet]
    charts = ws._charts

    # Component 1: At least one chart exists in the 'Marketing Data' sheet (0.4 pts)
    # This FAILS on initial_env (no charts) and PASSES on golden_env (1 chart)
    try:
        if len(charts) >= 1:
            print(f"PASS: Component 1 — Chart exists in '{target_sheet}' sheet ({len(charts)} chart(s) found) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — No charts found in '{target_sheet}' sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check charts: {e}")

    # Component 2: The chart is a column chart (BarChart with type='col') (0.3 pts)
    # A column chart is represented as BarChart with type='col' in openpyxl.
    # This FAILS on initial_env (no charts) and PASSES on golden_env (BarChart type=col).
    try:
        if len(charts) >= 1:
            from openpyxl.chart import BarChart
            chart = charts[0]
            is_bar_chart = isinstance(chart, BarChart)
            chart_type = getattr(chart, 'type', None)
            if is_bar_chart and chart_type == 'col':
                print(f"PASS: Component 2 — Chart is a column chart (BarChart type='col') (0.3 pts)")
                total_score += 0.3
            else:
                actual_type = f"{type(chart).__name__} type={chart_type}"
                print(f"FAIL: Component 2 — Expected BarChart type='col', found: {actual_type}")
        else:
            print("FAIL: Component 2 — No charts present to check type")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check chart type: {e}")

    # Component 3: Chart data references the marketing spend table (col B, rows 2-9) (0.3 pts)
    # The marketing spend table is in rows 1-9 (col A=Channel, col B=Monthly Spend).
    # The chart series data must reference column B (Monthly Spend), NOT column E (Conversion %).
    # This FAILS on initial_env (no charts) and PASSES on golden_env (series refs B2:B9).
    try:
        if len(charts) >= 1:
            chart = charts[0]
            series_list = chart.series
            found_spend_ref = False
            for s in series_list:
                try:
                    val_ref = s.val.numRef.f if (s.val and s.val.numRef) else ''
                    # Accept references to column B in rows 2-9 (marketing spend data)
                    # The reference should contain $B$ or column B rows within 1-9 range
                    # and must NOT reference column E (conversion rate table)
                    val_ref_upper = val_ref.upper()
                    if '$B$' in val_ref_upper and ('$B$2' in val_ref_upper or '$B$1' in val_ref_upper):
                        # Further check: the reference covers the marketing spend rows (B2:B9)
                        # and does not include column E data
                        if '$E$' not in val_ref_upper:
                            found_spend_ref = True
                            print(f"PASS: Component 3 — Chart data references marketing spend table: '{val_ref}' (0.3 pts)")
                            total_score += 0.3
                            break
                except Exception as inner_e:
                    print(f"  WARN: Could not parse series val ref: {inner_e}")

            if not found_spend_ref:
                # Print what we found for debugging
                refs_found = []
                for s in series_list:
                    try:
                        val_ref = s.val.numRef.f if (s.val and s.val.numRef) else 'N/A'
                        refs_found.append(val_ref)
                    except Exception:
                        refs_found.append('unknown')
                print(f"FAIL: Component 3 — Chart data does not reference marketing spend table (col B rows 2-9). Found refs: {refs_found}")
        else:
            print("FAIL: Component 3 — No charts present to check data reference")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check chart data reference: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
