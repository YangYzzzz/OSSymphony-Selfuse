"""
Initial Setup: Create a spreadsheet with a marketing spend table and conversion rate table
Task ID: osworld_calc_dual_chart_separate_tables_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_dual_chart_separate_tables_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Marketing Data"

    # --- Table 1: Marketing Spend (rows 1-9) ---
    # Headers in row 1
    ws.cell(row=1, column=1, value="Channel")
    ws.cell(row=1, column=2, value="Monthly Spend")

    # Style headers bold
    ws.cell(row=1, column=1).font = Font(bold=True)
    ws.cell(row=1, column=2).font = Font(bold=True)

    # Marketing channel data (rows 2-9 = 8 channels)
    marketing_data = [
        ("Social Media",    42500),
        ("Search (PPC)",    58300),
        ("Email Campaigns", 18750),
        ("Content/SEO",     23400),
        ("Display Ads",     31200),
        ("Influencer",      15600),
        ("Podcast Ads",      9800),
        ("Affiliate",       12300),
    ]

    for r, (channel, spend) in enumerate(marketing_data, 2):
        ws.cell(row=r, column=1, value=channel)
        ws.cell(row=r, column=2, value=spend)
        ws.cell(row=r, column=2).number_format = '#,##0'

    # Set column widths for Table 1
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16

    # --- Table 2: Conversion Rate (rows 12-20) ---
    # Headers in row 12
    ws.cell(row=12, column=4, value="Channel")
    ws.cell(row=12, column=5, value="Conversion %")

    # Style headers bold
    ws.cell(row=12, column=4).font = Font(bold=True)
    ws.cell(row=12, column=5).font = Font(bold=True)

    # Conversion rate data (rows 13-20 = 8 channels)
    conversion_data = [
        ("Social Media",    3.2),
        ("Search (PPC)",    5.8),
        ("Email Campaigns", 7.4),
        ("Content/SEO",     4.1),
        ("Display Ads",     1.9),
        ("Influencer",      2.7),
        ("Podcast Ads",     1.5),
        ("Affiliate",       4.6),
    ]

    for r, (channel, rate) in enumerate(conversion_data, 13):
        ws.cell(row=r, column=4, value=channel)
        ws.cell(row=r, column=5, value=rate)
        ws.cell(row=r, column=5).number_format = '0.0%'

    # Set column widths for Table 2
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 16

    # NOTE: NO charts in initial state — the task is to CREATE a chart from marketing spend table

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
