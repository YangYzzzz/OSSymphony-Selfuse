"""
Initial Setup: Create carrier performance comparison spreadsheet with delivery data
Task ID: calc_ops_052
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Deliveries ---
    ws = wb.active
    ws.title = 'Deliveries'

    # Headers
    ws['A1'] = 'Shipment'
    ws['B1'] = 'Carrier'
    ws['C1'] = 'Status'

    # Delivery data (10 shipments)
    deliveries = [
        ('S01', 'FedEx', 'On Time'),
        ('S02', 'UPS', 'Late'),
        ('S03', 'FedEx', 'On Time'),
        ('S04', 'DHL', 'On Time'),
        ('S05', 'UPS', 'On Time'),
        ('S06', 'FedEx', 'Late'),
        ('S07', 'DHL', 'On Time'),
        ('S08', 'UPS', 'On Time'),
        ('S09', 'FedEx', 'On Time'),
        ('S10', 'DHL', 'Late'),
    ]

    for r, (shipment, carrier, status) in enumerate(deliveries, 2):
        ws.cell(row=r, column=1, value=shipment)
        ws.cell(row=r, column=2, value=carrier)
        ws.cell(row=r, column=3, value=status)

    # Summary section headers (columns E-H)
    ws['E1'] = 'Carrier'
    ws['F1'] = 'Total'
    ws['G1'] = 'On Time'
    ws['H1'] = 'On Time %'

    # Carrier names in summary section
    ws['E2'] = 'FedEx'
    ws['E3'] = 'UPS'
    ws['E4'] = 'DHL'

    # F2:H4 intentionally LEFT EMPTY - the agent must fill these with formulas

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
