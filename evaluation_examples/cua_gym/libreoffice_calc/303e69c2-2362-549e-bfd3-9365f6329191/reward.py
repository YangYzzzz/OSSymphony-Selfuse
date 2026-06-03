"""
Reward Script: Carrier performance comparison with COUNTIFS and on-time percentage
Task ID: calc_ops_052
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): F2:F4 contain COUNTIFS formulas for total deliveries per carrier
  Component 2 (0.35): G2:G4 contain COUNTIFS formulas for on-time deliveries per carrier
  Component 3 (0.30): H2:H4 contain division formulas (G/F) for on-time percentage
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_052'


def is_countifs_formula(value, carrier_col="B", criteria_col=None, row_ref=None):
    """Check if value is a COUNTIFS formula referencing the carrier column."""
    if not isinstance(value, str):
        return False
    normalized = value.upper().replace(" ", "")
    # Must start with =COUNTIFS(
    if not normalized.startswith("=COUNTIFS("):
        return False
    # Must reference the B column (carrier column) somewhere
    if "B:" not in normalized and "B$" not in normalized and "$B" not in normalized and "B2" not in normalized:
        # Check for B column references more broadly
        if "B" not in normalized:
            return False
    return True


def is_division_formula(value, row):
    """Check if value is a division formula like =G{row}/F{row}."""
    if not isinstance(value, str):
        return False
    normalized = value.upper().replace(" ", "")
    # Must contain a division operator
    if "/" not in normalized:
        return False
    # Must start with =
    if not normalized.startswith("="):
        return False
    # Should reference G and F columns for this row
    g_ref = f"G{row}"
    f_ref = f"F{row}"
    if g_ref in normalized and f_ref in normalized:
        return True
    # Also accept generic patterns like =G2/F2 style
    if re.search(r'G\d+\s*/\s*F\d+', normalized):
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Deliveries' sheet exists
    if 'Deliveries' not in wb.sheetnames:
        print("CRITICAL: 'Deliveries' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Deliveries']

    # Component 1: F2:F4 contain COUNTIFS formulas for total deliveries per carrier (0.35 points)
    # Ground truth: FedEx total=4, UPS total=3, DHL total=3
    try:
        comp1_pass = 0
        for row in range(2, 5):
            cell_val = ws.cell(row=row, column=6).value  # Column F
            carrier = ws.cell(row=row, column=5).value    # Column E
            if is_countifs_formula(cell_val):
                print(f"  PASS: F{row} has COUNTIFS formula for {carrier}: {cell_val}")
                comp1_pass += 1
            else:
                print(f"  FAIL: F{row} expected COUNTIFS formula for {carrier}, found: {repr(cell_val)}")

        if comp1_pass == 3:
            print(f"PASS: Component 1 -- All 3 total count formulas present (0.35 pts)")
            total_score += 0.35
        elif comp1_pass >= 1:
            partial = round(0.35 * comp1_pass / 3, 2)
            print(f"PARTIAL: Component 1 -- {comp1_pass}/3 total count formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No COUNTIFS formulas found in F2:F4")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: G2:G4 contain COUNTIFS formulas for on-time deliveries per carrier (0.35 points)
    # Ground truth: FedEx on-time=3, UPS on-time=2, DHL on-time=2
    try:
        comp2_pass = 0
        for row in range(2, 5):
            cell_val = ws.cell(row=row, column=7).value  # Column G
            carrier = ws.cell(row=row, column=5).value    # Column E
            if is_countifs_formula(cell_val):
                # Additionally check it references "On Time" or the status column
                normalized = cell_val.upper().replace(" ", "")
                has_status_ref = ("ONTIME" in normalized or '"ON' in normalized
                                  or "C:" in normalized or "C$" in normalized)
                if has_status_ref:
                    print(f"  PASS: G{row} has COUNTIFS formula with on-time criteria for {carrier}: {cell_val}")
                    comp2_pass += 1
                else:
                    print(f"  PARTIAL: G{row} has COUNTIFS but missing on-time criteria for {carrier}: {cell_val}")
                    comp2_pass += 0.5
            else:
                print(f"  FAIL: G{row} expected COUNTIFS formula for {carrier}, found: {repr(cell_val)}")

        if comp2_pass >= 3:
            print(f"PASS: Component 2 -- All 3 on-time count formulas present (0.35 pts)")
            total_score += 0.35
        elif comp2_pass >= 1:
            partial = round(0.35 * comp2_pass / 3, 2)
            print(f"PARTIAL: Component 2 -- {comp2_pass}/3 on-time count formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No on-time COUNTIFS formulas found in G2:G4")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: H2:H4 contain division formulas for on-time percentage (0.30 points)
    # Ground truth: FedEx=75%, UPS=66.67%, DHL=66.67%
    try:
        comp3_pass = 0
        for row in range(2, 5):
            cell_val = ws.cell(row=row, column=8).value  # Column H
            carrier = ws.cell(row=row, column=5).value    # Column E
            if is_division_formula(cell_val, row):
                print(f"  PASS: H{row} has division formula for {carrier}: {cell_val}")
                comp3_pass += 1
            else:
                print(f"  FAIL: H{row} expected division formula for {carrier}, found: {repr(cell_val)}")

        if comp3_pass == 3:
            print(f"PASS: Component 3 -- All 3 percentage formulas present (0.30 pts)")
            total_score += 0.30
        elif comp3_pass >= 1:
            partial = round(0.30 * comp3_pass / 3, 2)
            print(f"PARTIAL: Component 3 -- {comp3_pass}/3 percentage formulas present ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No division formulas found in H2:H4")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
