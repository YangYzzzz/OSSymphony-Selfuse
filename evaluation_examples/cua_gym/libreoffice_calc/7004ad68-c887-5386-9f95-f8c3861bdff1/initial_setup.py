"""
Initial Setup: Subscription box business model spreadsheet
Task ID: calc_wf_092
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    int_fmt = '#,##0'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    def style_header(ws, row, max_col):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    # ========== Sheet 1: Pricing ==========
    ws_pricing = wb.active
    ws_pricing.title = "Pricing"

    # --- Box Tiers Section ---
    ws_pricing["A1"] = "Subscription Box Pricing"
    ws_pricing["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F5496")

    headers_tier = ["Tier", "Monthly Price", "Item Cost per Box", "Profit per Box", "Description"]
    for c, h in enumerate(headers_tier, 1):
        ws_pricing.cell(row=3, column=c, value=h)
    style_header(ws_pricing, 3, len(headers_tier))

    tier_data = [
        ["Basic", 29.99, 12.00, 17.99, "5 curated items, standard packaging"],
        ["Premium", 49.99, 22.00, 27.99, "8 premium items, gift-grade packaging"],
        ["Deluxe", 79.99, 35.00, 44.99, "12 luxury items, premium box with ribbon"],
    ]
    for r, row_data in enumerate(tier_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws_pricing.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (2, 3, 4):
                cell.number_format = currency_fmt

    # --- Shipping Costs Section ---
    ws_pricing["A8"] = "Shipping Costs by Region"
    ws_pricing["A8"].font = Font(name="Calibri", size=12, bold=True, color="2F5496")

    ship_headers = ["Region", "Cost per Box", "Estimated Delivery (days)", "Notes"]
    for c, h in enumerate(ship_headers, 1):
        ws_pricing.cell(row=10, column=c, value=h)
    style_header(ws_pricing, 10, len(ship_headers))

    ship_data = [
        ["Domestic (Continental US)", 5.00, "3-5", "USPS Priority Mail"],
        ["Domestic (Alaska/Hawaii)", 8.50, "5-7", "USPS Priority Mail"],
        ["Canada", 12.00, "7-10", "USPS First Class International"],
        ["International (Standard)", 15.00, "10-14", "USPS International Priority"],
        ["International (Express)", 25.00, "5-7", "FedEx International Express"],
    ]
    for r, row_data in enumerate(ship_data, 11):
        for c, val in enumerate(row_data, 1):
            cell = ws_pricing.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 2:
                cell.number_format = currency_fmt

    # --- Growth Assumptions Section ---
    ws_pricing["A18"] = "Growth Assumptions"
    ws_pricing["A18"].font = Font(name="Calibri", size=12, bold=True, color="2F5496")

    assumptions = [
        ["Starting Subscribers", 100],
        ["Monthly Growth Rate", 0.15],
        ["Monthly Churn Rate", 0.05],
        ["Tier Split - Basic", 0.70],
        ["Tier Split - Premium", 0.20],
        ["Tier Split - Deluxe", 0.10],
        ["Fixed Monthly Costs", 3000.00],
        ["Domestic Shipping %", 0.80],
        ["International Shipping %", 0.20],
    ]
    for c, h in enumerate(["Parameter", "Value"], 1):
        ws_pricing.cell(row=20, column=c, value=h)
    style_header(ws_pricing, 20, 2)

    for r, (param, val) in enumerate(assumptions, 21):
        cell_p = ws_pricing.cell(row=r, column=1, value=param)
        cell_v = ws_pricing.cell(row=r, column=2, value=val)
        cell_p.border = thin_border
        cell_v.border = thin_border
        if isinstance(val, float) and val < 1:
            cell_v.number_format = pct_fmt
        elif isinstance(val, float):
            cell_v.number_format = currency_fmt
        elif isinstance(val, int) and val > 10:
            cell_v.number_format = int_fmt

    # Column widths
    ws_pricing.column_dimensions["A"].width = 30
    ws_pricing.column_dimensions["B"].width = 16
    ws_pricing.column_dimensions["C"].width = 20
    ws_pricing.column_dimensions["D"].width = 18
    ws_pricing.column_dimensions["E"].width = 35

    # ========== Sheet 2: Projections ==========
    ws_proj = wb.create_sheet("Projections")

    ws_proj["A1"] = "12-Month Subscriber & Revenue Projections"
    ws_proj["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F5496")

    proj_headers = [
        "Month", "Subscribers", "New Subscribers", "Churned",
        "Revenue", "COGS", "Shipping Costs", "Fixed Costs",
        "Monthly Profit", "Cumulative Profit"
    ]
    for c, h in enumerate(proj_headers, 1):
        ws_proj.cell(row=3, column=c, value=h)
    style_header(ws_proj, 3, len(proj_headers))

    # Only populate Month column (1-12) - leave calculation columns empty
    # This is the initial state BEFORE the agent adds formulas
    months = [
        "Month 1", "Month 2", "Month 3", "Month 4",
        "Month 5", "Month 6", "Month 7", "Month 8",
        "Month 9", "Month 10", "Month 11", "Month 12"
    ]
    for r, m in enumerate(months, 4):
        cell = ws_proj.cell(row=r, column=1, value=m)
        cell.border = thin_border
        # Add borders to empty cells too
        for c in range(2, len(proj_headers) + 1):
            ws_proj.cell(row=r, column=c).border = thin_border

    # Column widths
    ws_proj.column_dimensions["A"].width = 12
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_proj.column_dimensions[col_letter].width = 16

    # ========== Sheet 3: Analysis ==========
    ws_analysis = wb.create_sheet("Analysis")

    ws_analysis["A1"] = "Break-Even Analysis"
    ws_analysis["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F5496")

    # Section headers only - no data or formulas
    ws_analysis["A3"] = "Break-Even Summary"
    ws_analysis["A3"].font = Font(name="Calibri", size=12, bold=True)

    be_labels = [
        "Monthly Profit Break-Even",
        "Subscribers at Monthly Break-Even",
        "Cumulative Profit Break-Even",
        "Final Cumulative Profit (Month 12)",
    ]
    for r, label in enumerate(be_labels, 5):
        cell = ws_analysis.cell(row=r, column=1, value=label)
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=11)
        ws_analysis.cell(row=r, column=2).border = thin_border

    ws_analysis["A11"] = "Charts"
    ws_analysis["A11"].font = Font(name="Calibri", size=12, bold=True)
    ws_analysis["A12"] = "(Charts to be added: Subscriber Growth, Revenue vs Costs)"
    ws_analysis["A12"].font = Font(name="Calibri", size=10, italic=True, color="808080")

    ws_analysis.column_dimensions["A"].width = 35
    ws_analysis.column_dimensions["B"].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
