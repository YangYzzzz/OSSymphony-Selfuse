"""
Reward Script: Subscription Box Business Model
Task ID: calc_wf_092
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Projections - subscriber data populated (B4=100, B5-B15 formulas)
  Component 2 (0.25): Projections - financial columns populated (Revenue, COGS, Shipping, Profit, Cumulative)
  Component 3 (0.20): Analysis - break-even data filled in (B5-B8 non-empty)
  Component 4 (0.15): Analysis - line chart for subscriber growth exists
  Component 5 (0.15): Analysis - area/stacked chart for revenue vs costs exists
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_092'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---- Precondition: Required sheets exist ----
    required_sheets = ['Projections', 'Analysis']
    for s in required_sheets:
        if s not in wb.sheetnames:
            print(f"CRITICAL: Missing required sheet '{s}'")
            print("REWARD: 0.0")
            return 0.0

    ws_proj = wb['Projections']
    ws_anal = wb['Analysis']

    # Component 1: Projections subscriber data populated (0.25 points)
    # Initial state: B4-B15 are all None. Golden state: B4=100, B5-B15 have formulas.
    try:
        b4_val = ws_proj['B4'].value
        # B4 should be 100 (starting subscribers)
        b4_ok = False
        if b4_val is not None:
            try:
                b4_ok = abs(float(b4_val) - 100) < 1
            except (ValueError, TypeError):
                # Could be a formula referencing pricing sheet
                b4_ok = isinstance(b4_val, str) and b4_val.startswith('=')

        # B5-B15 should have formulas or numeric values (growth model)
        formula_count = 0
        for row in range(5, 16):
            cell_val = ws_proj.cell(row=row, column=2).value
            if cell_val is not None:
                if isinstance(cell_val, str) and '=' in cell_val:
                    formula_count += 1
                elif isinstance(cell_val, (int, float)):
                    formula_count += 1

        if b4_ok and formula_count >= 8:
            print(f"PASS: Component 1 — Subscriber data populated. B4={b4_val}, {formula_count}/11 months have values (0.25 pts)")
            total_score += 0.25
        elif b4_ok or formula_count >= 4:
            partial = 0.125
            print(f"PARTIAL: Component 1 — B4 ok={b4_ok}, formulas={formula_count}/11 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — B4={b4_val} (ok={b4_ok}), formula_count={formula_count}/11")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Financial columns populated (0.25 points)
    # Initial state: columns E-J in rows 4-15 are all None.
    # Golden state: E (Revenue), F (COGS), G (Shipping), I (Profit), J (Cumulative) have formulas.
    try:
        # Check columns E, F, G, I, J for rows 4-15
        col_checks = {
            'E': 5,   # Revenue
            'F': 6,   # COGS
            'G': 7,   # Shipping
            'I': 9,   # Monthly Profit
            'J': 10,  # Cumulative Profit
        }
        filled_cols = 0
        for col_name, col_idx in col_checks.items():
            col_filled = 0
            for row in range(4, 16):
                val = ws_proj.cell(row=row, column=col_idx).value
                if val is not None:
                    col_filled += 1
            if col_filled >= 8:
                filled_cols += 1

        if filled_cols >= 4:
            print(f"PASS: Component 2 — {filled_cols}/5 financial columns populated (0.25 pts)")
            total_score += 0.25
        elif filled_cols >= 2:
            partial = 0.125
            print(f"PARTIAL: Component 2 — {filled_cols}/5 financial columns populated ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {filled_cols}/5 financial columns populated")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Analysis break-even data filled (0.20 points)
    # Initial state: B5-B8 are all None. Golden state: all have values.
    try:
        be_cells = ['B5', 'B6', 'B7', 'B8']
        filled_count = 0
        for coord in be_cells:
            val = ws_anal[coord].value
            if val is not None:
                filled_count += 1

        if filled_count >= 3:
            print(f"PASS: Component 3 — Break-even analysis data: {filled_count}/4 cells filled (0.20 pts)")
            total_score += 0.20
        elif filled_count >= 1:
            partial = 0.10
            print(f"PARTIAL: Component 3 — Break-even analysis data: {filled_count}/4 cells filled ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Break-even analysis data: 0/4 cells filled")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Line chart for subscriber growth (0.15 points)
    # Initial state: 0 charts. Golden state: at least one line chart.
    try:
        charts = ws_anal._charts
        line_chart_found = False
        for ch in charts:
            chart_type_name = type(ch).__name__
            if 'Line' in chart_type_name:
                line_chart_found = True
                # Verify it references subscriber data (column B of Projections)
                for s in ch.series:
                    ref = getattr(s, 'val', None) or getattr(s, 'values', None)
                    if ref is not None:
                        ref_str = str(ref)
                        if 'B' in ref_str or 'Subscribers' in ref_str:
                            print(f"PASS: Component 4 — Line chart found referencing subscriber data (0.15 pts)")
                            total_score += 0.15
                            break
                else:
                    # Line chart exists but couldn't confirm subscriber reference - still award
                    print(f"PASS: Component 4 — Line chart found (0.15 pts)")
                    total_score += 0.15
                break

        if not line_chart_found:
            # Check if any chart on any sheet could be the subscriber chart
            all_charts = []
            for sn in wb.sheetnames:
                all_charts.extend(wb[sn]._charts)
            for ch in all_charts:
                if 'Line' in type(ch).__name__:
                    line_chart_found = True
                    print(f"PASS: Component 4 — Line chart found on another sheet (0.15 pts)")
                    total_score += 0.15
                    break

        if not line_chart_found:
            print(f"FAIL: Component 4 — No line chart found for subscriber growth")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Area/stacked chart for revenue vs costs (0.15 points)
    # Initial state: 0 charts. Golden state: an AreaChart with multiple series.
    try:
        charts = ws_anal._charts
        area_chart_found = False
        for ch in charts:
            chart_type_name = type(ch).__name__
            if 'Area' in chart_type_name or ('Bar' in chart_type_name and len(ch.series) >= 2):
                if len(ch.series) >= 2:
                    area_chart_found = True
                    print(f"PASS: Component 5 — Revenue vs Costs chart found ({chart_type_name}, {len(ch.series)} series) (0.15 pts)")
                    total_score += 0.15
                    break

        if not area_chart_found:
            # Check other sheets too
            for sn in wb.sheetnames:
                for ch in wb[sn]._charts:
                    ctn = type(ch).__name__
                    if ('Area' in ctn or 'Bar' in ctn) and len(ch.series) >= 2:
                        area_chart_found = True
                        print(f"PASS: Component 5 — Revenue vs Costs chart found on sheet '{sn}' (0.15 pts)")
                        total_score += 0.15
                        break
                if area_chart_found:
                    break

        if not area_chart_found:
            print(f"FAIL: Component 5 — No area/stacked chart found for revenue vs costs")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
