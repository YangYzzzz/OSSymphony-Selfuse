"""
Initial Setup: Call center log with 500 rows for pivot table analysis
Task ID: calc_pivot_052
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

AGENTS = ['Agent1', 'Agent2', 'Agent3', 'Agent4', 'Agent5', 'Agent6', 'Agent7', 'Agent8']
CALL_TYPES = ['Sales', 'Support', 'Billing', 'Complaint']
RESOLUTIONS = ['Resolved', 'Escalated', 'Callback']

# Duration means per (agent, call_type) — tuned so Agent1/Sales~8.5, Agent1/Support~12.3, grand~11.2
DURATION_PARAMS = {
    ('Agent1', 'Sales'):      (8.5,  2.0),
    ('Agent1', 'Support'):    (12.3, 3.0),
    ('Agent1', 'Billing'):    (10.0, 2.5),
    ('Agent1', 'Complaint'):  (14.0, 3.5),
    ('Agent2', 'Sales'):      (9.2,  2.2),
    ('Agent2', 'Support'):    (11.8, 3.0),
    ('Agent2', 'Billing'):    (10.5, 2.8),
    ('Agent2', 'Complaint'):  (13.5, 3.2),
    ('Agent3', 'Sales'):      (7.8,  2.0),
    ('Agent3', 'Support'):    (13.0, 3.5),
    ('Agent3', 'Billing'):    (9.5,  2.5),
    ('Agent3', 'Complaint'):  (15.0, 4.0),
    ('Agent4', 'Sales'):      (10.0, 2.5),
    ('Agent4', 'Support'):    (11.5, 2.8),
    ('Agent4', 'Billing'):    (11.0, 3.0),
    ('Agent4', 'Complaint'):  (12.5, 3.0),
    ('Agent5', 'Sales'):      (8.0,  2.0),
    ('Agent5', 'Support'):    (12.0, 3.0),
    ('Agent5', 'Billing'):    (10.8, 2.5),
    ('Agent5', 'Complaint'):  (14.5, 3.5),
    ('Agent6', 'Sales'):      (9.5,  2.5),
    ('Agent6', 'Support'):    (11.0, 2.5),
    ('Agent6', 'Billing'):    (10.2, 2.8),
    ('Agent6', 'Complaint'):  (13.0, 3.0),
    ('Agent7', 'Sales'):      (8.8,  2.2),
    ('Agent7', 'Support'):    (12.5, 3.2),
    ('Agent7', 'Billing'):    (9.8,  2.5),
    ('Agent7', 'Complaint'):  (14.2, 3.5),
    ('Agent8', 'Sales'):      (9.0,  2.0),
    ('Agent8', 'Support'):    (11.2, 2.8),
    ('Agent8', 'Billing'):    (10.5, 2.5),
    ('Agent8', 'Complaint'):  (13.8, 3.2),
}

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_data(n=500):
    """Generate n rows of call center data with controlled duration distributions."""
    random.seed(42)
    rows = []
    base_date = datetime(2025, 1, 1)

    for i in range(1, n + 1):
        agent = random.choice(AGENTS)
        call_type = random.choice(CALL_TYPES)
        resolution = random.choice(RESOLUTIONS)
        date = base_date + timedelta(days=random.randint(0, 364))

        mean, std = DURATION_PARAMS[(agent, call_type)]
        duration = round(max(1, min(45, random.gauss(mean, std))), 1)

        rows.append((i, date, agent, call_type, duration, resolution))

    return rows


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CallLog'

    # Headers
    headers = ['CallID', 'Date', 'Agent', 'CallType', 'Duration', 'Resolution']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data
    data = generate_data(500)
    for r, (call_id, date, agent, call_type, duration, resolution) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=call_id)
        c_date = ws.cell(row=r, column=2, value=date)
        c_date.number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=3, value=agent)
        ws.cell(row=r, column=4, value=call_type)
        ws.cell(row=r, column=5, value=duration)
        ws.cell(row=r, column=6, value=resolution)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Compute and print actual averages for verification
    from collections import defaultdict
    sums = defaultdict(float)
    counts = defaultdict(int)
    for _, _, agent, call_type, duration, _ in data:
        sums[(agent, call_type)] += duration
        counts[(agent, call_type)] += 1
        sums[('TOTAL',)] += duration
        counts[('TOTAL',)] += 1

    print(f"Agent1/Sales avg: {sums[('Agent1','Sales')]/counts[('Agent1','Sales')]:.1f} (n={counts[('Agent1','Sales')]})")
    print(f"Agent1/Support avg: {sums[('Agent1','Support')]/counts[('Agent1','Support')]:.1f} (n={counts[('Agent1','Support')]})")
    print(f"Grand total avg: {sums[('TOTAL',)]/counts[('TOTAL',)]:.1f}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
