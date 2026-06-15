"""
Reward Script: Pivot table analyzing call center data — average call duration by agent and call type
Task ID: calc_pivot_052
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): A new sheet exists (beyond CallLog) containing pivot-style data
  Component 2 (0.25): Row headers contain all 8 agents (Agent1-Agent8)
  Component 3 (0.25): Column headers contain call types (Sales, Support, Billing, Complaint)
  Component 4 (0.25): Data values are reasonable averages and Grand Total row exists
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_052'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: CallLog sheet must still exist (data integrity gate)
    if 'CallLog' not in wb.sheetnames:
        print("FAIL: CallLog sheet missing — data integrity compromised")
        print("REWARD: 0.0")
        return 0.0

    # Find the pivot sheet — any sheet OTHER than CallLog
    pivot_sheets = [name for name in wb.sheetnames if name != 'CallLog']

    # Component 1: A new pivot sheet exists (0.25 points)
    # This FAILS on initial (only CallLog exists) and PASSES on golden (Pivot sheet added)
    try:
        if len(pivot_sheets) >= 1:
            pivot_name = pivot_sheets[0]
            ps = wb[pivot_name]
            # Verify it has meaningful content (at least a few rows and columns)
            if ps.max_row >= 5 and ps.max_column >= 3:
                print(f"PASS: Component 1 — Pivot sheet '{pivot_name}' exists with {ps.max_row} rows, {ps.max_column} cols (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — Sheet '{pivot_name}' exists but too small: {ps.max_row} rows, {ps.max_column} cols")
        else:
            print("FAIL: Component 1 — No pivot sheet found (only CallLog exists)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Early exit if no pivot sheet
    if len(pivot_sheets) < 1:
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ps = wb[pivot_sheets[0]]

    # Collect all cell values from the pivot sheet for analysis
    all_values = []
    for row in ps.iter_rows(min_row=1, max_row=ps.max_row, max_col=ps.max_column, values_only=True):
        all_values.append(list(row))

    # Flatten all string values for searching
    all_str_values = set()
    for row in all_values:
        for v in row:
            if v is not None:
                all_str_values.add(str(v).strip())

    # Component 2: Row headers contain Agent1 through Agent8 (0.25 points)
    # This FAILS on initial (no pivot sheet) and PASSES on golden
    try:
        expected_agents = [f'Agent{i}' for i in range(1, 9)]
        found_agents = []
        for agent in expected_agents:
            if agent in all_str_values:
                found_agents.append(agent)

        agent_ratio = len(found_agents) / len(expected_agents)
        if agent_ratio >= 0.75:  # at least 6 of 8 agents
            points = 0.25 * agent_ratio
            print(f"PASS: Component 2 — Found {len(found_agents)}/8 agents as row headers ({points:.3f} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 2 — Only found {len(found_agents)}/8 agents: {found_agents}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column headers contain call types (Sales, Support, Billing, Complaint) (0.25 points)
    # This FAILS on initial (no pivot sheet) and PASSES on golden
    try:
        expected_types = ['Sales', 'Support', 'Billing', 'Complaint']
        found_types = []
        for ct in expected_types:
            if ct in all_str_values:
                found_types.append(ct)

        type_ratio = len(found_types) / len(expected_types)
        if type_ratio >= 0.75:  # at least 3 of 4 call types
            points = 0.25 * type_ratio
            print(f"PASS: Component 3 — Found {len(found_types)}/4 call types as column headers ({points:.3f} pts)")
            total_score += points
        else:
            print(f"FAIL: Component 3 — Only found {len(found_types)}/4 call types: {found_types}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data contains reasonable average values and Grand Total row (0.25 points)
    # Average call durations should be between 1 and 45 (the Duration range from context)
    # This FAILS on initial (no pivot sheet) and PASSES on golden
    try:
        numeric_values = []
        grand_total_matches = []

        for row in all_values:
            for v in row:
                if isinstance(v, (int, float)) and v != 0:
                    numeric_values.append(v)
                if v is not None:
                    lowered = str(v).strip().lower()
                    if lowered in ('grand total', 'total', 'grand_total', 'grandtotal'):
                        grand_total_matches.append(str(v))

        # Check that numeric values are in reasonable average range (1-45)
        reasonable_count = sum(1 for v in numeric_values if 1.0 <= v <= 45.0)

        if len(numeric_values) >= 10 and reasonable_count >= len(numeric_values) * 0.8:
            total_score += 0.15
            print(f"PASS: Component 4a — {len(numeric_values)} numeric values, {reasonable_count} in reasonable range (0.15 pts)")
        else:
            print(f"FAIL: Component 4a — Found {len(numeric_values)} numeric values, {reasonable_count} in range 1-45")

        if len(grand_total_matches) > 0:
            total_score += 0.10
            print(f"PASS: Component 4b — Grand Total row found: {grand_total_matches[0]} (0.10 pts)")
        else:
            print(f"FAIL: Component 4b — No Grand Total row found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
