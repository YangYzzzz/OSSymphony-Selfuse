"""
Initial Setup: Break-even and margin model for Goal Seek task
Task ID: calc_gen_goalseek_047
Domain: libreoffice_calc

Creates a spreadsheet with cost inputs for a margin model.
C5 (Selling Price) is empty — the agent will use Goal Seek to fill it.
C6 (Gross Margin formula) is empty — agent adds it.
Rows 9-20 are empty — agent builds the sensitivity table.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_goalseek_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MarginModel'

    # --- Title area ---
    ws['A1'] = 'Product Pricing & Margin Analysis'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # --- Input section label ---
    ws['A2'] = 'Input Parameters'
    ws['A2'].font = Font(name='Calibri', size=11, bold=True, italic=True)

    # --- Cost inputs ---
    # B2: COGS per unit, C2: 28.50
    ws['B2'] = 'COGS per unit'
    ws['C2'] = 28.50
    ws['C2'].number_format = '$#,##0.00'

    # B3: Fixed overhead per unit, C3: 8.25
    ws['B3'] = 'Fixed overhead per unit'
    ws['C3'] = 8.25
    ws['C3'].number_format = '$#,##0.00'

    # B4: Target Margin %, C4: 0.35
    ws['B4'] = 'Target Margin %'
    ws['C4'] = 0.35
    ws['C4'].number_format = '0%'

    # B5: Selling Price — EMPTY (Goal Seek target cell)
    ws['B5'] = 'Selling Price'
    # C5 intentionally left empty — agent uses Goal Seek to populate it

    # B6: Gross Margin — EMPTY formula (agent adds the formula)
    ws['B6'] = 'Gross Margin'
    # C6 intentionally left empty — agent adds =(C5-C2-C3)/C5

    # --- Separator row ---
    ws['A7'] = ''
    ws['B7'] = 'Notes'
    ws['C7'] = 'Use Goal Seek to find the selling price that yields target margin'
    ws['C7'].font = Font(name='Calibri', size=10, italic=True, color='595959')

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14

    # --- Style the input labels ---
    label_font = Font(name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    for row in range(2, 7):
        cell_b = ws.cell(row=row, column=2)
        cell_b.font = label_font
        cell_b.fill = header_fill

    # Rows 9-20 are intentionally left empty (agent will build sensitivity table here)
    # The task says "Rows 9-19 are empty (for sensitivity table)" in the initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Initial state:')
    print('  - MarginModel sheet with cost inputs')
    print('  - C2=28.50 (COGS), C3=8.25 (Fixed overhead), C4=0.35 (Target margin)')
    print('  - C5 empty (Selling Price - Goal Seek target)')
    print('  - C6 empty (Gross Margin formula - to be added)')
    print('  - Rows 9-20 empty (sensitivity table to be built)')


create_initial()
