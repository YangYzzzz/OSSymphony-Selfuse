"""
Reward Script: Break-even/margin model with Goal Seek and sensitivity table
Task ID: calc_gen_goalseek_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): C5 contains Goal Seek result ~56.54 (selling price for 35% margin)
  Component 2 (0.20): C6 contains gross margin formula =(C5-C2-C3)/C5
  Component 3 (0.10): Sensitivity table headers A9='Price', B9='Margin %'
  Component 4 (0.20): A10:A20 contain price series 50,60,...,150
  Component 5 (0.10): B10:B20 contain margin formulas referencing each row's price
  Component 6 (0.10): Conditional formatting on B10:B20 (green >= 0.35, red < 0.35)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_goalseek_047'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check MarginModel sheet exists
    if 'MarginModel' not in wb.sheetnames:
        print("FAIL: Sheet 'MarginModel' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['MarginModel']

    # Component 1: Goal Seek result in C5 — selling price for 35% margin (0.30 points)
    # COGS=28.5, Fixed overhead=8.25, Target margin=35%
    # Required price = (28.5 + 8.25) / (1 - 0.35) = 36.75 / 0.65 = 56.538...
    try:
        c5_val = ws['C5'].value
        expected_price = (28.5 + 8.25) / (1 - 0.35)  # = 56.538...
        if c5_val is not None:
            try:
                c5_float = float(c5_val)
                if abs(c5_float - expected_price) <= 0.5:
                    print(f"PASS: Component 1 — C5 contains Goal Seek price {c5_float:.4f} (expected ~{expected_price:.4f}) (0.30 pts)")
                    total_score += 0.30
                else:
                    print(f"FAIL: Component 1 — C5 = {c5_float}, expected ~{expected_price:.4f} (within 0.5 tolerance)")
            except (ValueError, TypeError):
                print(f"FAIL: Component 1 — C5 value '{c5_val}' is not numeric")
        else:
            print("FAIL: Component 1 — C5 is empty (Goal Seek not applied)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Gross margin formula in C6 = =(C5-C2-C3)/C5 (0.20 points)
    # The formula must reference C5, C2, and C3 in a margin calculation
    try:
        c6_val = ws['C6'].value
        if c6_val is not None and isinstance(c6_val, str):
            # Normalize: strip spaces, uppercase
            formula_norm = c6_val.upper().replace(' ', '')
            # Accept any formula that computes (C5-C2-C3)/C5 or equivalent
            if ('C5' in formula_norm and 'C2' in formula_norm and 'C3' in formula_norm and
                    formula_norm.startswith('=')):
                # Check it has division structure: (C5-...) / C5
                if '/C5' in formula_norm or '/ C5' in c6_val.upper():
                    print(f"PASS: Component 2 — C6 contains margin formula: {c6_val} (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 2 — C6 formula '{c6_val}' missing /C5 division")
            else:
                print(f"FAIL: Component 2 — C6 formula '{c6_val}' doesn't match expected pattern =(C5-C2-C3)/C5")
        elif c6_val is None:
            print("FAIL: Component 2 — C6 is empty (no gross margin formula)")
        else:
            print(f"FAIL: Component 2 — C6 value '{c6_val}' is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sensitivity table headers A9='Price', B9='Margin %' (0.10 points)
    try:
        a9_val = ws['A9'].value
        b9_val = ws['B9'].value
        a9_ok = a9_val is not None and str(a9_val).strip().lower() == 'price'
        b9_ok = b9_val is not None and 'margin' in str(b9_val).strip().lower()
        if a9_ok and b9_ok:
            print(f"PASS: Component 3 — Sensitivity table headers: A9='{a9_val}', B9='{b9_val}' (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Headers missing or wrong: A9={repr(a9_val)}, B9={repr(b9_val)} (expected 'Price' and 'Margin %')")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Price series A10:A20 = 50, 60, 70, ..., 150 in $10 increments (0.20 points)
    try:
        expected_prices = list(range(50, 160, 10))  # [50, 60, 70, ..., 150]
        actual_prices = []
        for row in range(10, 21):
            val = ws.cell(row=row, column=1).value
            actual_prices.append(val)

        # Check all 11 price points exist
        price_matches = 0
        for i, (expected, actual) in enumerate(zip(expected_prices, actual_prices)):
            if actual is not None:
                try:
                    if abs(float(actual) - expected) <= 0.01:
                        price_matches += 1
                except (ValueError, TypeError):
                    pass

        if price_matches == 11:
            print(f"PASS: Component 4 — All 11 price points ($50-$150) present in A10:A20 (0.20 pts)")
            total_score += 0.20
        elif price_matches >= 6:
            print(f"PARTIAL: Component 4 — {price_matches}/11 price points correct in A10:A20 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Only {price_matches}/11 price points correct. Got: {actual_prices}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Margin formulas in B10:B20 (0.10 points)
    # Each cell should have a formula like =(A10-$C$2-$C$3)/A10
    try:
        formula_count = 0
        for row in range(10, 21):
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                # Formula should reference the row's column A and the cost cells
                val_norm = val.upper().replace(' ', '')
                if ('C2' in val_norm or '$C$2' in val_norm) and ('C3' in val_norm or '$C$3' in val_norm):
                    formula_count += 1

        if formula_count >= 11:
            print(f"PASS: Component 5 — All 11 margin formulas present in B10:B20 (0.10 pts)")
            total_score += 0.10
        elif formula_count >= 6:
            print(f"PARTIAL: Component 5 — {formula_count}/11 margin formulas found in B10:B20 (0.05 pts)")
            total_score += 0.05
        else:
            # Also check if they are stored as static numeric values (data table result)
            numeric_count = 0
            for row in range(10, 21):
                val = ws.cell(row=row, column=2).value
                if val is not None:
                    try:
                        float_val = float(val)
                        # Validate the value is a reasonable margin (between -2 and 1)
                        if -2.0 <= float_val <= 1.0:
                            numeric_count += 1
                    except (ValueError, TypeError):
                        pass
            if numeric_count >= 11:
                print(f"PASS: Component 5 — All 11 margin values present as numbers in B10:B20 (0.10 pts)")
                total_score += 0.10
            elif numeric_count >= 6:
                print(f"PARTIAL: Component 5 — {numeric_count}/11 margin values in B10:B20 (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — Only {formula_count} formulas and {numeric_count} numeric values in B10:B20")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on B10:B20 (green >= 0.35, red < 0.35) (0.10 points)
    try:
        cf_rules = ws.conditional_formatting
        target_range_found = False
        has_green_rule = False
        has_red_rule = False

        for cf in cf_rules:
            cf_str = str(cf)
            # Check if range covers B10:B20 or similar range
            if 'B10' in cf_str and 'B20' in cf_str:
                target_range_found = True
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', '')
                    operator = getattr(rule, 'operator', '')
                    formula = getattr(rule, 'formula', [])

                    # Check for green >= 0.35 rule
                    if (rule_type == 'cellIs' and
                            operator in ('greaterThanOrEqual', 'greaterThan') and
                            formula and
                            '0.35' in str(formula)):
                        has_green_rule = True

                    # Check for red < 0.35 rule
                    if (rule_type == 'cellIs' and
                            operator == 'lessThan' and
                            formula and
                            '0.35' in str(formula)):
                        has_red_rule = True

        if target_range_found and has_green_rule and has_red_rule:
            print("PASS: Component 6 — Conditional formatting on B10:B20 with green (>=35%) and red (<35%) rules (0.10 pts)")
            total_score += 0.10
        elif target_range_found and (has_green_rule or has_red_rule):
            print(f"PARTIAL: Component 6 — CF range found but missing rules: green={has_green_rule}, red={has_red_rule} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Conditional formatting on B10:B20 not found or incomplete: range={target_range_found}, green={has_green_rule}, red={has_red_rule}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
