"""
Initial Setup: Reconciliation spreadsheet with financial comparison data
Task ID: calc_gg5_047
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()

    # --- Differences Sheet ---
    ws_diff = wb.active
    ws_diff.title = "Differences"

    # Headers
    ws_diff.cell(row=1, column=1, value="System Value")
    ws_diff.cell(row=1, column=2, value="Bank Value")

    # Style headers
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    header_align = Alignment(horizontal="center")
    for col in [1, 2]:
        cell = ws_diff.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws_diff.column_dimensions["A"].width = 18
    ws_diff.column_dimensions["B"].width = 18

    # Generate 150 rows of financial data
    # About 40% will have mismatches (roughly 60 mismatched rows)
    for r in range(2, 152):
        # Generate a base system value - realistic financial amounts
        base_val = round(random.uniform(100.00, 25000.00), 2)
        system_val = base_val

        # Decide if this row matches or not
        if random.random() < 0.40:
            # Mismatch: bank value differs by a small amount
            diff = round(random.uniform(0.01, 500.00), 2)
            if random.random() < 0.5:
                diff = -diff
            bank_val = round(system_val + diff, 2)
        else:
            # Match: identical values
            bank_val = system_val

        ws_diff.cell(row=r, column=1, value=system_val)
        ws_diff.cell(row=r, column=2, value=bank_val)

        # Number format for currency-like appearance
        ws_diff.cell(row=r, column=1).number_format = '#,##0.00'
        ws_diff.cell(row=r, column=2).number_format = '#,##0.00'

    # --- Report Sheet ---
    ws_report = wb.create_sheet("Report")

    # Headers only - no data (agent must fill via macro)
    report_headers = [
        ("A1", "Metric"),
        ("B1", "Value"),
    ]
    for coord, val in report_headers:
        cell = ws_report[coord]
        cell.value = val
        cell.font = Font(name="Calibri", size=11, bold=True)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # Metric labels in column A (rows 2-5) but NO values in B
    ws_report["A2"] = "Total Rows"
    ws_report["A3"] = "Matched"
    ws_report["A4"] = "Mismatched"
    ws_report["A5"] = "Discrepancy Rows"

    ws_report.column_dimensions["A"].width = 20
    ws_report.column_dimensions["B"].width = 60

    # Save
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
