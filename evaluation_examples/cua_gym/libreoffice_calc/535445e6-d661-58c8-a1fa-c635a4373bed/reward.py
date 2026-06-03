"""
Reward Script: Reconciliation macro verification
Task ID: calc_gg5_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Yellow highlighting on all mismatched rows (and only those)
  Component 2 (0.1): Report sheet Total Rows = 150
  Component 3 (0.1): Report sheet Matched count correct
  Component 4 (0.1): Report sheet Mismatched count correct
  Component 5 (0.3): Report sheet Discrepancy Rows list correct
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_047'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition: required sheets exist ---
    if 'Differences' not in wb.sheetnames:
        print("FAIL: 'Differences' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'Report' not in wb.sheetnames:
        print("FAIL: 'Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_diff = wb['Differences']
    ws_report = wb['Report']

    # --- First, independently compute the ground truth from the data ---
    expected_mismatches = []
    expected_matches = 0
    for row_num in range(2, 152):
        a_val = ws_diff.cell(row=row_num, column=1).value
        b_val = ws_diff.cell(row=row_num, column=2).value
        if a_val != b_val:
            expected_mismatches.append(row_num)
        else:
            expected_matches += 1
    expected_mismatch_count = len(expected_mismatches)
    expected_total = 150

    # Component 1: Yellow highlighting on mismatched rows (0.4 points)
    # This checks that ALL rows where A != B are highlighted yellow,
    # AND that matched rows are NOT highlighted yellow.
    try:
        correctly_highlighted = 0
        incorrectly_highlighted = 0
        missing_highlight = 0

        for row_num in range(2, 152):
            a_val = ws_diff.cell(row=row_num, column=1).value
            b_val = ws_diff.cell(row=row_num, column=2).value
            is_mismatch = (a_val != b_val)

            # Check if either column A or B in this row has yellow fill
            has_yellow = False
            for col in [1, 2]:
                cell = ws_diff.cell(row=row_num, column=col)
                try:
                    fg_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                    if fg_rgb and cell.fill.patternType == 'solid':
                        # Yellow = FFFFFF00 or FFFF00 variants
                        rgb_str = str(fg_rgb).upper()
                        # Check for yellow: high R, high G, low B
                        # Common yellow codes: FFFFFF00, FFFFFF00
                        if rgb_str in ('FFFFFF00', '00FFFF00') or 'FFFF00' in rgb_str:
                            has_yellow = True
                            break
                except Exception:
                    pass

            if is_mismatch and has_yellow:
                correctly_highlighted += 1
            elif is_mismatch and not has_yellow:
                missing_highlight += 1
            elif not is_mismatch and has_yellow:
                incorrectly_highlighted += 1

        total_mismatch_rows = len(expected_mismatches)
        if total_mismatch_rows > 0 and missing_highlight == 0 and incorrectly_highlighted == 0:
            print(f"PASS: Component 1 — All {correctly_highlighted} mismatched rows highlighted yellow, "
                  f"no false positives (0.4 pts)")
            total_score += 0.4
        elif total_mismatch_rows > 0 and correctly_highlighted > 0:
            # Partial credit: proportion of correctly highlighted minus penalty for false positives
            ratio = correctly_highlighted / total_mismatch_rows
            penalty = incorrectly_highlighted / 150.0
            partial = max(0.0, 0.4 * (ratio - penalty))
            print(f"PARTIAL: Component 1 — {correctly_highlighted}/{total_mismatch_rows} mismatched rows "
                  f"highlighted, {incorrectly_highlighted} false positives ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No mismatched rows highlighted yellow "
                  f"(missing={missing_highlight}, false_pos={incorrectly_highlighted})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Report sheet Total Rows = 150 (0.1 points)
    try:
        total_rows_val = ws_report.cell(row=2, column=2).value
        if total_rows_val is not None:
            try:
                total_rows_num = int(float(str(total_rows_val)))
            except (ValueError, TypeError):
                total_rows_num = None
            if total_rows_num == expected_total:
                print(f"PASS: Component 2 — Total Rows = {total_rows_num} (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 2 — Total Rows = {total_rows_val}, expected {expected_total}")
        else:
            print(f"FAIL: Component 2 — Total Rows cell is empty")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Report sheet Matched count (0.1 points)
    try:
        matched_val = ws_report.cell(row=3, column=2).value
        if matched_val is not None:
            try:
                matched_num = int(float(str(matched_val)))
            except (ValueError, TypeError):
                matched_num = None
            if matched_num == expected_matches:
                print(f"PASS: Component 3 — Matched = {matched_num} (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Matched = {matched_val}, expected {expected_matches}")
        else:
            print(f"FAIL: Component 3 — Matched cell is empty")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Report sheet Mismatched count (0.1 points)
    try:
        mismatch_val = ws_report.cell(row=4, column=2).value
        if mismatch_val is not None:
            try:
                mismatch_num = int(float(str(mismatch_val)))
            except (ValueError, TypeError):
                mismatch_num = None
            if mismatch_num == expected_mismatch_count:
                print(f"PASS: Component 4 — Mismatched = {mismatch_num} (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 4 — Mismatched = {mismatch_val}, expected {expected_mismatch_count}")
        else:
            print(f"FAIL: Component 4 — Mismatched cell is empty")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Report sheet Discrepancy Rows list correct (0.3 points)
    try:
        disc_val = ws_report.cell(row=5, column=2).value
        if disc_val is not None:
            disc_str = str(disc_val)
            # Extract all numbers from the discrepancy string
            found_rows = sorted([int(x) for x in re.findall(r'\d+', disc_str)])
            expected_row_list = sorted(expected_mismatches)

            if found_rows == expected_row_list:
                print(f"PASS: Component 5 — Discrepancy rows list matches exactly "
                      f"({len(found_rows)} rows) (0.3 pts)")
                total_score += 0.3
            elif len(found_rows) > 0:
                # Partial credit based on overlap
                found_set = set(found_rows)
                expected_set = set(expected_row_list)
                correct = len(found_set & expected_set)
                total_expected = len(expected_set)
                extra = len(found_set - expected_set)
                if total_expected > 0:
                    precision = correct / len(found_set) if len(found_set) > 0 else 0
                    recall = correct / total_expected
                    partial = 0.3 * min(precision, recall)
                    print(f"PARTIAL: Component 5 — {correct}/{total_expected} correct rows listed, "
                          f"{extra} extra ({partial:.3f} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 5 — Could not match discrepancy rows")
            else:
                print(f"FAIL: Component 5 — No row numbers found in discrepancy cell: '{disc_str}'")
        else:
            print(f"FAIL: Component 5 — Discrepancy Rows cell is empty")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
