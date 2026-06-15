"""
Initial Setup: Demand forecast comparison sheet with raw data only.
Task ID: calc_ops_046
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Forecast ---
    ws = wb.active
    ws.title = 'Forecast'

    # Headers in row 1
    headers = ['Month', 'Actual', 'Simple Avg', 'Weighted MA', 'Exp Smooth (a=0.3)']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Data rows 2-7: Month, Actual only. Columns C, D, E left empty.
    data = [
        ['Jan', 100],
        ['Feb', 120],
        ['Mar', 110],
        ['Apr', 130],
        ['May', 125],
        ['Jun', 140],
    ]

    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for r, (month, actual) in enumerate(data, 2):
        cell_a = ws.cell(row=r, column=1, value=month)
        cell_a.alignment = Alignment(horizontal="center")
        cell_a.border = data_border

        cell_b = ws.cell(row=r, column=2, value=actual)
        cell_b.number_format = '0'
        cell_b.alignment = Alignment(horizontal="center")
        cell_b.border = data_border

        # Leave C, D, E empty but with borders for visual clarity
        for col in range(3, 6):
            c = ws.cell(row=r, column=col)
            c.border = data_border
            c.alignment = Alignment(horizontal="center")
            c.number_format = '0.00'

    # Row 8: empty separator (no content)

    # Row 9: MAD label in A9, C9/D9/E9 left empty for formulas
    cell_mad = ws.cell(row=9, column=1, value='MAD')
    cell_mad.font = Font(bold=True, size=11, name='Calibri')
    cell_mad.alignment = Alignment(horizontal="center")
    cell_mad.border = data_border

    for col in range(2, 6):
        c = ws.cell(row=9, column=col)
        c.border = data_border
        c.alignment = Alignment(horizontal="center")
        c.number_format = '0.00'

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 22

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
