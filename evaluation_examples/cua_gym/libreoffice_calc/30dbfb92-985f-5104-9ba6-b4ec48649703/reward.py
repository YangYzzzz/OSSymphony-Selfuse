"""
Reward Script: Demand forecast comparison sheet with 3 methods + MAD
Task ID: calc_ops_046
Domain: libreoffice_calc
Scoring:
  - Component 1: Exp Smoothing seed E2=100 (0.10)
  - Component 2: Simple Average formulas C3:C7 (0.20)
  - Component 3: Exponential Smoothing formulas E3:E7 (0.20)
  - Component 4: Weighted MA formulas D5:D7 (0.20)
  - Component 5: MAD formulas C9, D9, E9 (0.30)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_046'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def formula_matches(actual, pattern):
    """Check if actual formula matches expected pattern (regex or exact)."""
    actual_norm = normalize_formula(actual)
    pattern_norm = pattern.upper().replace(' ', '')
    return bool(re.match(pattern_norm, actual_norm))


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Forecast' sheet must exist
    if 'Forecast' not in wb.sheetnames:
        print("CRITICAL: 'Forecast' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Forecast']

    # Component 1: Exponential Smoothing seed value E2=100 (0.10 points)
    # In initial_env, E2 is empty. In golden_env, E2 should be 100.
    try:
        e2_val = ws['E2'].value
        if e2_val is not None and float(e2_val) == 100:
            print(f"PASS: Component 1 - E2 seed value = {e2_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 - Expected E2=100, found: {e2_val}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Simple Average formulas in C3:C7 (0.20 points)
    # C3=AVERAGE(B2:B2), C4=AVERAGE(B2:B3), C5=AVERAGE(B2:B4), C6=AVERAGE(B2:B5), C7=AVERAGE(B2:B6)
    try:
        simple_avg_expected = {
            'C3': '=AVERAGE(B2:B2)',
            'C4': '=AVERAGE(B2:B3)',
            'C5': '=AVERAGE(B2:B4)',
            'C6': '=AVERAGE(B2:B5)',
            'C7': '=AVERAGE(B2:B6)',
        }
        sa_pass = 0
        for coord, expected in simple_avg_expected.items():
            actual = ws[coord].value
            if actual is not None and normalize_formula(actual) == normalize_formula(expected):
                sa_pass += 1
            else:
                print(f"  DETAIL: {coord} expected {expected}, found {repr(actual)}")
        if sa_pass == 5:
            print(f"PASS: Component 2 - All 5 Simple Avg formulas correct (0.20 pts)")
            total_score += 0.20
        elif sa_pass > 0:
            sa_score = (sa_pass / 5) * 0.20
            print(f"PARTIAL: Component 2 - {sa_pass}/5 Simple Avg formulas correct ({sa_score:.2f} pts)")
            total_score += sa_score
        else:
            print(f"FAIL: Component 2 - No Simple Avg formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Exponential Smoothing formulas in E3:E7 (0.20 points)
    # E3=E2+0.3*(B2-E2), E4=E3+0.3*(B3-E3), etc.
    try:
        exp_smooth_expected = {
            'E3': '=E2+0.3*(B2-E2)',
            'E4': '=E3+0.3*(B3-E3)',
            'E5': '=E4+0.3*(B4-E4)',
            'E6': '=E5+0.3*(B5-E5)',
            'E7': '=E6+0.3*(B6-E6)',
        }
        es_pass = 0
        for coord, expected in exp_smooth_expected.items():
            actual = ws[coord].value
            if actual is not None and normalize_formula(actual) == normalize_formula(expected):
                es_pass += 1
            else:
                print(f"  DETAIL: {coord} expected {expected}, found {repr(actual)}")
        if es_pass == 5:
            print(f"PASS: Component 3 - All 5 Exp Smoothing formulas correct (0.20 pts)")
            total_score += 0.20
        elif es_pass > 0:
            es_score = (es_pass / 5) * 0.20
            print(f"PARTIAL: Component 3 - {es_pass}/5 Exp Smoothing formulas correct ({es_score:.2f} pts)")
            total_score += es_score
        else:
            print(f"FAIL: Component 3 - No Exp Smoothing formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Weighted MA formulas in D5:D7 (0.20 points)
    # Weights: 0.5 (most recent), 0.3, 0.2
    # D5=0.5*B4+0.3*B3+0.2*B2, D6=0.5*B5+0.3*B4+0.2*B3, D7=0.5*B6+0.3*B5+0.2*B4
    try:
        wma_expected = {
            'D5': '=0.5*B4+0.3*B3+0.2*B2',
            'D6': '=0.5*B5+0.3*B4+0.2*B3',
            'D7': '=0.5*B6+0.3*B5+0.2*B4',
        }
        wma_pass = 0
        for coord, expected in wma_expected.items():
            actual = ws[coord].value
            if actual is not None and normalize_formula(actual) == normalize_formula(expected):
                wma_pass += 1
            else:
                print(f"  DETAIL: {coord} expected {expected}, found {repr(actual)}")
        if wma_pass == 3:
            print(f"PASS: Component 4 - All 3 Weighted MA formulas correct (0.20 pts)")
            total_score += 0.20
        elif wma_pass > 0:
            wma_score = (wma_pass / 3) * 0.20
            print(f"PARTIAL: Component 4 - {wma_pass}/3 Weighted MA formulas correct ({wma_score:.2f} pts)")
            total_score += wma_score
        else:
            print(f"FAIL: Component 4 - No Weighted MA formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: MAD formulas in C9, D9, E9 (0.30 points, 0.10 each)
    # C9: MAD for Simple Avg over months where forecasts exist (Feb-Jun = 5 months)
    # D9: MAD for Weighted MA over months where forecasts exist (Apr-Jun = 3 months)
    # E9: MAD for Exp Smooth over all months (Jan-Jun = 6 months)
    try:
        mad_expected = {
            'C9': '=(ABS(B3-C3)+ABS(B4-C4)+ABS(B5-C5)+ABS(B6-C6)+ABS(B7-C7))/5',
            'D9': '=(ABS(B5-D5)+ABS(B6-D6)+ABS(B7-D7))/3',
            'E9': '=(ABS(B2-E2)+ABS(B3-E3)+ABS(B4-E4)+ABS(B5-E5)+ABS(B6-E6)+ABS(B7-E7))/6',
        }
        mad_pass = 0
        for coord, expected in mad_expected.items():
            actual = ws[coord].value
            if actual is not None and normalize_formula(actual) == normalize_formula(expected):
                mad_pass += 1
            else:
                print(f"  DETAIL: {coord} expected {expected}, found {repr(actual)}")
        if mad_pass == 3:
            print(f"PASS: Component 5 - All 3 MAD formulas correct (0.30 pts)")
            total_score += 0.30
        elif mad_pass > 0:
            mad_score = (mad_pass / 3) * 0.30
            print(f"PARTIAL: Component 5 - {mad_pass}/3 MAD formulas correct ({mad_score:.2f} pts)")
            total_score += mad_score
        else:
            print(f"FAIL: Component 5 - No MAD formulas found")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
