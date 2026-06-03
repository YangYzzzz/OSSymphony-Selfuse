"""
Initial Setup: Create a Finance spreadsheet with Item and Amount columns
Task ID: calc_lf_073
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Finance ---
    ws = wb.active
    ws.title = 'Finance'

    # Headers
    ws.cell(row=1, column=1, value='Item')
    ws.cell(row=1, column=2, value='Amount')

    # Data rows - financial line items with plain numeric values
    data = [
        ['Revenue', 150000],
        ['COGS', 90000],
        ['Gross Profit', 60000],
        ['Expenses', 35000],
        ['Net Income', 25000],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
