"""
Reward Script: Create CurrencyStyle named style and apply to B2:B6
Task ID: calc_lf_073
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): CurrencyStyle named style exists with correct properties
  Component 2 (0.3): B2:B6 have currency number format '$#,##0.00'
  Component 3 (0.2): B2:B6 are right-aligned
  Component 4 (0.2): B2:B6 have light green background (#C6EFCE)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_073'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Finance sheet must exist
    if 'Finance' not in wb.sheetnames:
        print("CRITICAL: 'Finance' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Finance']

    # Component 1: CurrencyStyle named style exists with correct properties (0.3 points)
    try:
        style_names = list(wb.named_styles)
        if 'CurrencyStyle' in style_names:
            # Find the style object
            style_obj = None
            for ns in wb._named_styles:
                if ns.name == 'CurrencyStyle':
                    style_obj = ns
                    break

            if style_obj is not None:
                sub_score = 0.0
                # Check number format
                if style_obj.number_format == '$#,##0.00':
                    sub_score += 0.1
                    print(f"PASS: CurrencyStyle number_format is '$#,##0.00'")
                else:
                    print(f"FAIL: CurrencyStyle number_format is '{style_obj.number_format}', expected '$#,##0.00'")

                # Check alignment
                if style_obj.alignment and style_obj.alignment.horizontal == 'right':
                    sub_score += 0.1
                    print(f"PASS: CurrencyStyle alignment is 'right'")
                else:
                    h = style_obj.alignment.horizontal if style_obj.alignment else None
                    print(f"FAIL: CurrencyStyle alignment.horizontal is '{h}', expected 'right'")

                # Check fill color
                try:
                    fg_rgb = style_obj.fill.fgColor.rgb if style_obj.fill.fgColor else None
                    if fg_rgb == 'FFC6EFCE' and style_obj.fill.patternType == 'solid':
                        sub_score += 0.1
                        print(f"PASS: CurrencyStyle fill is light green (FFC6EFCE)")
                    else:
                        print(f"FAIL: CurrencyStyle fill fgColor={fg_rgb}, patternType={style_obj.fill.patternType}, expected FFC6EFCE/solid")
                except Exception as e:
                    print(f"FAIL: CurrencyStyle fill check error: {e}")

                if sub_score > 0:
                    total_score += sub_score
                print(f"Component 1 subtotal: {sub_score}/0.3")
            else:
                print("FAIL: Component 1 -- CurrencyStyle found in names but not in _named_styles")
        else:
            print(f"FAIL: Component 1 -- 'CurrencyStyle' not in named_styles: {style_names}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: B2:B6 have currency number format '$#,##0.00' (0.3 points)
    try:
        cells_with_format = 0
        target_cells = ['B2', 'B3', 'B4', 'B5', 'B6']
        for coord in target_cells:
            cell = ws[coord]
            if cell.number_format == '$#,##0.00':
                cells_with_format += 1
            else:
                print(f"FAIL: {coord} number_format is '{cell.number_format}', expected '$#,##0.00'")

        if cells_with_format == 5:
            print(f"PASS: Component 2 -- All 5 cells (B2:B6) have '$#,##0.00' format (0.3 pts)")
            total_score += 0.3
        elif cells_with_format > 0:
            partial = round(0.3 * cells_with_format / 5, 2)
            print(f"PARTIAL: Component 2 -- {cells_with_format}/5 cells have correct format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No cells in B2:B6 have '$#,##0.00' format")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: B2:B6 are right-aligned (0.2 points)
    try:
        cells_aligned = 0
        for coord in target_cells:
            cell = ws[coord]
            if cell.alignment and cell.alignment.horizontal == 'right':
                cells_aligned += 1
            else:
                h = cell.alignment.horizontal if cell.alignment else None
                print(f"FAIL: {coord} alignment.horizontal is '{h}', expected 'right'")

        if cells_aligned == 5:
            print(f"PASS: Component 3 -- All 5 cells (B2:B6) are right-aligned (0.2 pts)")
            total_score += 0.2
        elif cells_aligned > 0:
            partial = round(0.2 * cells_aligned / 5, 2)
            print(f"PARTIAL: Component 3 -- {cells_aligned}/5 cells are right-aligned ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No cells in B2:B6 are right-aligned")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: B2:B6 have light green background FFC6EFCE (0.2 points)
    try:
        cells_filled = 0
        for coord in target_cells:
            cell = ws[coord]
            try:
                fg_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                fill_type = cell.fill.patternType
                if fg_rgb == 'FFC6EFCE' and fill_type == 'solid':
                    cells_filled += 1
                else:
                    print(f"FAIL: {coord} fill fgColor={fg_rgb}, patternType={fill_type}, expected FFC6EFCE/solid")
            except Exception as e:
                print(f"FAIL: {coord} fill check error: {e}")

        if cells_filled == 5:
            print(f"PASS: Component 4 -- All 5 cells (B2:B6) have light green background (0.2 pts)")
            total_score += 0.2
        elif cells_filled > 0:
            partial = round(0.2 * cells_filled / 5, 2)
            print(f"PARTIAL: Component 4 -- {cells_filled}/5 cells have correct fill ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No cells in B2:B6 have light green background")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
