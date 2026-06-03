"""
Reward Script: Build comprehensive claim analysis pivot in Sheet2
Task ID: osworld_calc_pivot_count_invoice_012
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Sheet2 has pivot structure (header row + 4 claim type rows) — 0.30 pts
  Component 2: Pivot counts are correct (North/South/East/West counts per claim type) — 0.40 pts
  Component 3: Grand Total column present and correct (row totals: Medical=17, Auto=12, Property=11, Life=8) — 0.10 pts
  Component 4: Percentage-of-total column present and correct — 0.10 pts
  Component 5: Rows sorted by total count descending (Medical > Auto > Property > Life) — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_012'

# Expected pivot values derived from the raw data in Sheet1
# 48 total claims: Medical=17, Auto=12, Property=11, Life=8
EXPECTED_COUNTS = {
    'Medical':  {'North': 4, 'South': 5, 'East': 4, 'West': 4, 'Grand Total': 17},
    'Auto':     {'North': 3, 'South': 3, 'East': 3, 'West': 3, 'Grand Total': 12},
    'Property': {'North': 3, 'South': 2, 'East': 3, 'West': 3, 'Grand Total': 11},
    'Life':     {'North': 2, 'South': 2, 'East': 2, 'West': 2, 'Grand Total': 8},
}
EXPECTED_PERCENTAGES = {
    'Medical': 35.42,
    'Auto': 25.0,
    'Property': 22.92,
    'Life': 16.67,
}
# Sorted order descending by total count
EXPECTED_ORDER = ['Medical', 'Auto', 'Property', 'Life']
TOTAL_CLAIMS = 48

REGIONS = ['North', 'South', 'East', 'West']


def find_pivot_in_sheet(ws):
    """
    Scan Sheet2 to find the pivot header row and data rows.
    Returns (header_row_idx, data_rows) where data_rows is a list of dicts:
      {'claim_type': str, 'North': int, 'South': int, 'East': int, 'West': int,
       'Grand Total': int_or_None, 'pct': float_or_None}
    Returns (None, []) if pivot structure is not found.
    """
    header_row_idx = None
    col_map = {}  # column name -> col index (1-based)

    # Scan for a header row that contains 'Claim Type' and at least one region
    for row_idx in range(1, ws.max_row + 1):
        row_vals = {}
        for col_idx in range(1, ws.max_column + 1):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                row_vals[col_idx] = str(val).strip()

        # Check if this row looks like a pivot header
        # Must have something that could be "Claim Type" plus at least one region
        row_str_vals = list(row_vals.values())
        has_claim_type = any(v.lower() in ('claim type', 'claimtype') for v in row_str_vals)
        has_region = any(v in ('North', 'South', 'East', 'West') for v in row_str_vals)

        if has_claim_type and has_region:
            header_row_idx = row_idx
            # Build col_map
            for col_idx, val in row_vals.items():
                col_map[val] = col_idx
                # Normalize Grand Total variants
                if val.lower() in ('grand total', 'grandtotal', 'total'):
                    col_map['Grand Total'] = col_idx
                # Normalize % variants
                if '%' in val.lower() or 'percent' in val.lower():
                    col_map['%'] = col_idx
            break

    if header_row_idx is None:
        return None, []

    # Extract data rows below the header
    data_rows = []
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        claim_type_col = col_map.get('Claim Type') or col_map.get('ClaimType')
        if claim_type_col is None:
            # Try to find by position (first non-empty column with a text value)
            for col_idx in range(1, ws.max_column + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val and isinstance(val, str) and val.strip():
                    claim_type_col = col_idx
                    break

        if claim_type_col is None:
            continue

        ct_val = ws.cell(row=row_idx, column=claim_type_col).value
        if not ct_val or not isinstance(ct_val, str):
            continue
        ct_val = ct_val.strip()

        # Skip grand total summary rows
        if ct_val.lower() in ('grand total', 'total', 'grandtotal'):
            continue

        row_data = {'claim_type': ct_val}
        for region in REGIONS:
            region_col = col_map.get(region)
            if region_col:
                val = ws.cell(row=row_idx, column=region_col).value
                try:
                    row_data[region] = int(val) if val is not None else None
                except (ValueError, TypeError):
                    row_data[region] = None
            else:
                row_data[region] = None

        # Grand total
        gt_col = col_map.get('Grand Total')
        if gt_col:
            val = ws.cell(row=row_idx, column=gt_col).value
            try:
                row_data['Grand Total'] = int(val) if val is not None else None
            except (ValueError, TypeError):
                row_data['Grand Total'] = None
        else:
            row_data['Grand Total'] = None

        # Percentage
        pct_col = col_map.get('%')
        if pct_col:
            val = ws.cell(row=row_idx, column=pct_col).value
            try:
                row_data['pct'] = float(val) if val is not None else None
            except (ValueError, TypeError):
                row_data['pct'] = None
        else:
            row_data['pct'] = None

        data_rows.append(row_data)

    return header_row_idx, data_rows


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws2 = wb['Sheet2']

    # Scan for pivot header and data rows
    header_row_idx, data_rows = find_pivot_in_sheet(ws2)

    # -----------------------------------------------------------------------
    # Component 1: Sheet2 has pivot structure — claim type rows present (0.30 pts)
    # Initial env: Sheet2 is empty → FAIL
    # Golden env: Sheet2 has 4 claim type rows → PASS
    # -----------------------------------------------------------------------
    try:
        found_types = {r['claim_type'] for r in data_rows}
        expected_types = set(EXPECTED_ORDER)
        found_expected = found_types & expected_types

        if header_row_idx is not None and len(found_expected) >= 4:
            print(f"PASS: Component 1 — Pivot structure present with all 4 claim types: {sorted(found_expected)} (0.30 pts)")
            total_score += 0.30
        elif header_row_idx is not None and len(found_expected) >= 2:
            # Partial: at least 2 types found
            partial = round(0.30 * len(found_expected) / 4, 2)
            print(f"PARTIAL: Component 1 — Pivot has {len(found_expected)}/4 claim types: {sorted(found_expected)} ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — No pivot structure found in Sheet2 (found claim types: {sorted(found_types)})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Pivot counts correct by region (0.40 pts)
    # Initial env: Sheet2 empty → FAIL
    # Golden env: Medical[N=4,S=5,E=4,W=4], Auto[3,3,3,3], Property[3,2,3,3], Life[2,2,2,2] → PASS
    # -----------------------------------------------------------------------
    try:
        if not data_rows:
            print("FAIL: Component 2 — No data rows found in pivot")
        else:
            # Build lookup by claim type
            row_by_type = {r['claim_type']: r for r in data_rows}
            correct_cells = 0
            total_cells = len(EXPECTED_COUNTS) * len(REGIONS)  # 4 types x 4 regions = 16

            for ctype, expected_regions in EXPECTED_COUNTS.items():
                if ctype not in row_by_type:
                    print(f"  MISS: Claim type '{ctype}' not found in pivot")
                    continue
                row = row_by_type[ctype]
                for region in REGIONS:
                    expected_val = expected_regions[region]
                    actual_val = row.get(region)
                    if actual_val == expected_val:
                        correct_cells += 1
                    else:
                        print(f"  FAIL: {ctype}/{region}: expected {expected_val}, found {actual_val}")

            pct_correct = correct_cells / total_cells
            comp2_score = round(0.40 * pct_correct, 4)
            if pct_correct >= 1.0:
                print(f"PASS: Component 2 — All {total_cells} regional counts correct (0.40 pts)")
                total_score += 0.40
            elif pct_correct >= 0.5:
                print(f"PARTIAL: Component 2 — {correct_cells}/{total_cells} regional counts correct ({comp2_score} pts)")
                total_score += comp2_score
            else:
                print(f"FAIL: Component 2 — Only {correct_cells}/{total_cells} regional counts correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Grand Total column correct (0.10 pts)
    # Initial env: Sheet2 empty → FAIL
    # Golden env: Medical=17, Auto=12, Property=11, Life=8 → PASS
    # -----------------------------------------------------------------------
    try:
        if not data_rows:
            print("FAIL: Component 3 — No data rows to check Grand Total")
        else:
            row_by_type = {r['claim_type']: r for r in data_rows}
            grand_totals_ok = 0
            for ctype, expected_regions in EXPECTED_COUNTS.items():
                expected_gt = expected_regions['Grand Total']
                if ctype in row_by_type:
                    actual_gt = row_by_type[ctype].get('Grand Total')
                    if actual_gt == expected_gt:
                        grand_totals_ok += 1
                    else:
                        print(f"  FAIL GT: {ctype} Grand Total expected {expected_gt}, found {actual_gt}")

            if grand_totals_ok >= 4:
                print(f"PASS: Component 3 — All Grand Total column values correct (Medical=17, Auto=12, Property=11, Life=8) (0.10 pts)")
                total_score += 0.10
            elif grand_totals_ok >= 2:
                partial = round(0.10 * grand_totals_ok / 4, 3)
                print(f"PARTIAL: Component 3 — {grand_totals_ok}/4 Grand Totals correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Grand Total column missing or incorrect ({grand_totals_ok}/4 correct)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Percentage-of-total column correct (0.10 pts)
    # Initial env: Sheet2 empty → FAIL
    # Golden env: Medical=35.42%, Auto=25.0%, Property=22.92%, Life=16.67% → PASS
    # -----------------------------------------------------------------------
    try:
        if not data_rows:
            print("FAIL: Component 4 — No data rows to check percentages")
        else:
            row_by_type = {r['claim_type']: r for r in data_rows}
            pct_ok = 0
            has_pct_col = any(r.get('pct') is not None for r in data_rows)

            if not has_pct_col:
                print("FAIL: Component 4 — Percentage-of-total column not found in pivot")
            else:
                for ctype, expected_pct in EXPECTED_PERCENTAGES.items():
                    if ctype in row_by_type:
                        actual_pct = row_by_type[ctype].get('pct')
                        if actual_pct is not None and abs(float(actual_pct) - expected_pct) <= 0.1:
                            pct_ok += 1
                        else:
                            print(f"  FAIL PCT: {ctype} % expected ~{expected_pct}, found {actual_pct}")

                if pct_ok >= 4:
                    print(f"PASS: Component 4 — All percentage values correct (Medical=35.42%, Auto=25%, Property=22.92%, Life=16.67%) (0.10 pts)")
                    total_score += 0.10
                elif pct_ok >= 2:
                    partial = round(0.10 * pct_ok / 4, 3)
                    print(f"PARTIAL: Component 4 — {pct_ok}/4 percentage values correct ({partial} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 4 — Percentage column incorrect ({pct_ok}/4 values match)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Rows sorted by total count descending (0.10 pts)
    # Initial env: Sheet2 empty → FAIL
    # Golden env: Medical(17) > Auto(12) > Property(11) > Life(8) → PASS
    # -----------------------------------------------------------------------
    try:
        if len(data_rows) < 4:
            print(f"FAIL: Component 5 — Not enough rows to check sort order ({len(data_rows)} found)")
        else:
            # Get actual order of claim types as they appear
            actual_order = [r['claim_type'] for r in data_rows if r['claim_type'] in EXPECTED_ORDER]

            # Check if sorted by Grand Total descending
            # Get grand totals in row order
            grand_totals_in_order = []
            for r in data_rows:
                if r['claim_type'] in EXPECTED_ORDER:
                    gt = r.get('Grand Total')
                    if gt is None:
                        # Fall back to sum of region counts
                        region_sum = sum(r.get(reg, 0) or 0 for reg in REGIONS)
                        gt = region_sum if region_sum > 0 else None
                    grand_totals_in_order.append((r['claim_type'], gt))

            # Verify descending order
            is_sorted_desc = all(
                grand_totals_in_order[i][1] >= grand_totals_in_order[i+1][1]
                for i in range(len(grand_totals_in_order) - 1)
                if grand_totals_in_order[i][1] is not None and grand_totals_in_order[i+1][1] is not None
            )

            # Also check exact expected order
            is_exact_order = actual_order[:4] == EXPECTED_ORDER

            if is_sorted_desc and len(grand_totals_in_order) >= 4:
                print(f"PASS: Component 5 — Rows sorted descending by total count: {actual_order[:4]} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 — Rows not sorted descending. Actual order: {actual_order}")
                print(f"  Expected: {EXPECTED_ORDER}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
