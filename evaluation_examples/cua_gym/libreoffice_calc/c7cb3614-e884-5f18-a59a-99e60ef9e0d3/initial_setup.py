"""
Initial Setup: Insurance Claims Spreadsheet
Task ID: osworld_calc_pivot_count_invoice_012
Domain: libreoffice_calc

Creates Sheet1 with insurance claim records (Claim ID, Claim Type, Region,
Claim Date, Claim Amount, Status) and a blank Sheet2.
The task requires the agent to build a pivot table in Sheet2 counting claims
by Claim Type x Region with percentage-of-total and sorted by total descending.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Claims ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Claim ID', 'Claim Type', 'Region', 'Claim Date', 'Claim Amount', 'Status']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 14
    ws1.column_dimensions['C'].width = 10
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 12

    # Realistic insurance claim data
    # 40 rows covering Medical/Auto/Property/Life x North/South/East/West
    data = [
        # Claim ID, Claim Type, Region, Claim Date, Claim Amount, Status
        ('CLM-001', 'Medical',  'North', '2024-01-08', 12450.00, 'Approved'),
        ('CLM-002', 'Auto',     'South', '2024-01-12', 8320.50,  'Approved'),
        ('CLM-003', 'Property', 'East',  '2024-01-15', 24500.00, 'Pending'),
        ('CLM-004', 'Life',     'West',  '2024-01-20', 150000.00,'Approved'),
        ('CLM-005', 'Medical',  'South', '2024-01-22', 5670.75,  'Approved'),
        ('CLM-006', 'Auto',     'North', '2024-01-25', 11200.00, 'Denied'),
        ('CLM-007', 'Property', 'West',  '2024-01-28', 38900.00, 'Approved'),
        ('CLM-008', 'Medical',  'East',  '2024-02-03', 9840.25,  'Pending'),
        ('CLM-009', 'Auto',     'East',  '2024-02-07', 6750.00,  'Approved'),
        ('CLM-010', 'Life',     'North', '2024-02-10', 200000.00,'Approved'),
        ('CLM-011', 'Medical',  'West',  '2024-02-14', 3200.50,  'Approved'),
        ('CLM-012', 'Property', 'South', '2024-02-18', 15600.00, 'Denied'),
        ('CLM-013', 'Auto',     'West',  '2024-02-21', 9100.75,  'Approved'),
        ('CLM-014', 'Medical',  'North', '2024-02-25', 18750.00, 'Approved'),
        ('CLM-015', 'Life',     'East',  '2024-03-01', 75000.00, 'Pending'),
        ('CLM-016', 'Property', 'North', '2024-03-05', 42300.00, 'Approved'),
        ('CLM-017', 'Medical',  'South', '2024-03-08', 7825.00,  'Approved'),
        ('CLM-018', 'Auto',     'North', '2024-03-12', 14600.00, 'Approved'),
        ('CLM-019', 'Medical',  'East',  '2024-03-15', 6310.25,  'Denied'),
        ('CLM-020', 'Property', 'West',  '2024-03-19', 29800.00, 'Approved'),
        ('CLM-021', 'Life',     'South', '2024-03-22', 125000.00,'Approved'),
        ('CLM-022', 'Auto',     'South', '2024-03-25', 4500.00,  'Approved'),
        ('CLM-023', 'Medical',  'West',  '2024-03-28', 11200.50, 'Pending'),
        ('CLM-024', 'Property', 'East',  '2024-04-02', 18400.00, 'Approved'),
        ('CLM-025', 'Medical',  'North', '2024-04-05', 8900.00,  'Approved'),
        ('CLM-026', 'Auto',     'East',  '2024-04-09', 7230.75,  'Approved'),
        ('CLM-027', 'Life',     'West',  '2024-04-12', 300000.00,'Denied'),
        ('CLM-028', 'Medical',  'South', '2024-04-16', 4150.25,  'Approved'),
        ('CLM-029', 'Property', 'North', '2024-04-19', 55000.00, 'Approved'),
        ('CLM-030', 'Auto',     'West',  '2024-04-22', 8750.00,  'Pending'),
        ('CLM-031', 'Medical',  'East',  '2024-04-25', 13450.00, 'Approved'),
        ('CLM-032', 'Life',     'North', '2024-04-29', 180000.00,'Approved'),
        ('CLM-033', 'Property', 'South', '2024-05-03', 22100.50, 'Approved'),
        ('CLM-034', 'Auto',     'North', '2024-05-06', 5980.00,  'Denied'),
        ('CLM-035', 'Medical',  'West',  '2024-05-10', 9640.75,  'Approved'),
        ('CLM-036', 'Property', 'West',  '2024-05-14', 34500.00, 'Approved'),
        ('CLM-037', 'Medical',  'South', '2024-05-17', 6720.00,  'Approved'),
        ('CLM-038', 'Auto',     'East',  '2024-05-20', 10250.50, 'Approved'),
        ('CLM-039', 'Life',     'South', '2024-05-24', 95000.00, 'Pending'),
        ('CLM-040', 'Property', 'East',  '2024-05-27', 17850.00, 'Approved'),
        ('CLM-041', 'Medical',  'North', '2024-06-01', 21300.00, 'Approved'),
        ('CLM-042', 'Auto',     'West',  '2024-06-04', 3890.25,  'Approved'),
        ('CLM-043', 'Medical',  'East',  '2024-06-08', 7560.50,  'Denied'),
        ('CLM-044', 'Life',     'East',  '2024-06-11', 240000.00,'Approved'),
        ('CLM-045', 'Property', 'North', '2024-06-15', 48700.00, 'Pending'),
        ('CLM-046', 'Medical',  'West',  '2024-06-18', 5200.00,  'Approved'),
        ('CLM-047', 'Auto',     'South', '2024-06-21', 12450.75, 'Approved'),
        ('CLM-048', 'Medical',  'South', '2024-06-25', 8900.25,  'Approved'),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Freeze header row
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Empty (agent will build pivot table here) ---
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'] = ''  # completely empty, ready for agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
