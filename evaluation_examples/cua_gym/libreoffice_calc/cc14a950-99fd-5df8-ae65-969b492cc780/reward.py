"""
Reward Script: Apply conditional number format to B2:B5
Task ID: calc_lf_060
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): B2 has the correct custom number format
  Component 2 (0.25): B3 has the correct custom number format
  Component 3 (0.25): B4 has the correct custom number format
  Component 4 (0.25): B5 has the correct custom number format
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_060'

EXPECTED_FORMAT = '[GREEN]#,##0.00;[RED](#,##0.00);[BLUE]0.00'

# Expected values (precondition check — not scored)
EXPECTED_VALUES = {
    'B2': 45000,
    'B3': -12500,
    'B4': 0,
    'B5': 8750.5,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet 'PnL' must exist
    if 'PnL' not in wb.sheetnames:
        print(f"FAIL: Sheet 'PnL' not found. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PnL']

    # Precondition gate: data values must be intact (not scored)
    for coord, expected_val in EXPECTED_VALUES.items():
        cell_val = ws[coord].value
        if cell_val is None:
            print(f"FAIL: Precondition — {coord} is empty, expected {expected_val}")
            print("REWARD: 0.0")
            return 0.0
        try:
            if abs(float(cell_val) - expected_val) > 0.01:
                print(f"FAIL: Precondition — {coord} value is {cell_val}, expected {expected_val}")
                print("REWARD: 0.0")
                return 0.0
        except (ValueError, TypeError):
            print(f"FAIL: Precondition — {coord} value {cell_val!r} is not numeric")
            print("REWARD: 0.0")
            return 0.0

    print("PASS: Precondition — all data values intact")

    # Component 1: B2 number format (0.25 points)
    # Task changes B2 from 'General' to the conditional format
    try:
        nf = ws['B2'].number_format
        if nf == EXPECTED_FORMAT:
            print(f"PASS: Component 1 — B2 number_format is correct (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — B2 number_format is {nf!r}, expected {EXPECTED_FORMAT!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 number format (0.25 points)
    try:
        nf = ws['B3'].number_format
        if nf == EXPECTED_FORMAT:
            print(f"PASS: Component 2 — B3 number_format is correct (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — B3 number_format is {nf!r}, expected {EXPECTED_FORMAT!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B4 number format (0.25 points)
    try:
        nf = ws['B4'].number_format
        if nf == EXPECTED_FORMAT:
            print(f"PASS: Component 3 — B4 number_format is correct (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — B4 number_format is {nf!r}, expected {EXPECTED_FORMAT!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: B5 number format (0.25 points)
    try:
        nf = ws['B5'].number_format
        if nf == EXPECTED_FORMAT:
            print(f"PASS: Component 4 — B5 number_format is correct (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — B5 number_format is {nf!r}, expected {EXPECTED_FORMAT!r}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
