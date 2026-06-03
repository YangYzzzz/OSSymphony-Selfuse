"""
Initial Setup: Create PnL sheet with financial data for conditional number format task
Task ID: calc_lf_060
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: PnL ---
    ws = wb.active
    ws.title = 'PnL'

    # Headers
    ws.cell(row=1, column=1, value='Item')
    ws.cell(row=1, column=2, value='Amount')

    # Data rows
    data = [
        ('Revenue', 45000),
        ('Loss', -12500),
        ('Break-even', 0),
        ('Profit', 8750.5),
    ]
    for r, (item, amount) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=item)
        ws.cell(row=r, column=2, value=amount)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
