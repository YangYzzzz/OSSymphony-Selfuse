"""
Initial Setup: Create Test_Matrix spreadsheet with test case data, no validation or formatting on Result column.
Task ID: calc_gcv_093
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test_Matrix"

    # --- Headers ---
    headers = ["Test Case", "Module", "Tester", "Result"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_side = Side(style="thin", color="000000")
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- 29 Test Case Rows (rows 2-30) ---
    modules = [
        "Authentication", "User Profile", "Dashboard", "Payments",
        "Notifications", "Search", "Reporting", "Admin Panel",
        "File Upload", "API Gateway", "Data Export", "Settings"
    ]
    testers = [
        "Sarah Chen", "Marcus Johnson", "Emily Park", "David Kim",
        "Rachel Martinez", "James Wilson", "Priya Sharma", "Alex Thompson",
        "Maria Garcia", "Kevin Liu"
    ]

    test_cases = [
        ("TC-001", "Authentication", "Login with valid credentials", "Sarah Chen"),
        ("TC-002", "Authentication", "Login with invalid password", "Sarah Chen"),
        ("TC-003", "Authentication", "Password reset flow", "Marcus Johnson"),
        ("TC-004", "User Profile", "Update display name", "Emily Park"),
        ("TC-005", "User Profile", "Change profile picture", "Emily Park"),
        ("TC-006", "User Profile", "Edit email address", "David Kim"),
        ("TC-007", "Dashboard", "Load performance metrics", "Rachel Martinez"),
        ("TC-008", "Dashboard", "Filter by date range", "Rachel Martinez"),
        ("TC-009", "Dashboard", "Export dashboard as PDF", "James Wilson"),
        ("TC-010", "Payments", "Process credit card payment", "Priya Sharma"),
        ("TC-011", "Payments", "Apply discount code", "Priya Sharma"),
        ("TC-012", "Payments", "Refund processing", "Alex Thompson"),
        ("TC-013", "Notifications", "Email notification delivery", "Maria Garcia"),
        ("TC-014", "Notifications", "Push notification settings", "Maria Garcia"),
        ("TC-015", "Notifications", "In-app notification badge", "Kevin Liu"),
        ("TC-016", "Search", "Full-text search accuracy", "Sarah Chen"),
        ("TC-017", "Search", "Search with filters", "Marcus Johnson"),
        ("TC-018", "Search", "Search result pagination", "Emily Park"),
        ("TC-019", "Reporting", "Generate monthly report", "David Kim"),
        ("TC-020", "Reporting", "Schedule automated reports", "Rachel Martinez"),
        ("TC-021", "Reporting", "Custom report builder", "James Wilson"),
        ("TC-022", "Admin Panel", "User management CRUD", "Priya Sharma"),
        ("TC-023", "Admin Panel", "Role-based access control", "Alex Thompson"),
        ("TC-024", "File Upload", "Upload CSV file", "Maria Garcia"),
        ("TC-025", "File Upload", "Drag-and-drop upload", "Kevin Liu"),
        ("TC-026", "API Gateway", "Rate limiting validation", "Sarah Chen"),
        ("TC-027", "API Gateway", "API key authentication", "Marcus Johnson"),
        ("TC-028", "Data Export", "Export to Excel format", "Emily Park"),
        ("TC-029", "Settings", "Toggle dark mode", "David Kim"),
    ]

    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r, (tc_id, module, description, tester) in enumerate(test_cases, 2):
        # Column A: Test Case (ID + description)
        cell_a = ws.cell(row=r, column=1, value=f"{tc_id}: {description}")
        cell_a.border = data_border

        # Column B: Module
        cell_b = ws.cell(row=r, column=2, value=module)
        cell_b.border = data_border

        # Column C: Tester
        cell_c = ws.cell(row=r, column=3, value=tester)
        cell_c.border = data_border

        # Column D: Result - intentionally empty (task is to add validation + formatting here)
        cell_d = ws.cell(row=r, column=4)
        cell_d.border = data_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
