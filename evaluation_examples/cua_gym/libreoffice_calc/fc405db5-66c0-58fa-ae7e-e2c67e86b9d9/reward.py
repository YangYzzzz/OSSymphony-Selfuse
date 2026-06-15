"""
Reward Script: Gantt Chart in LibreOffice Calc
Task ID: calc_grs_005
Domain: libreoffice_calc
Scoring:
  Component 1: End Date formulas in column D (0.20)
  Component 2: Date headers in row 1 spanning timeline (0.25)
  Component 3: Conditional formatting - blue task bars (0.25)
  Component 4: Conditional formatting - completion color (0.15)
  Component 5: Freeze panes set correctly (0.15)
"""

import os
import openpyxl
from datetime import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_005'


def verify_task(file_path):
    """
    Verify Gantt chart task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the Gantt chart sheet (accept any sheet name)
    ws = wb.worksheets[0]
    print(f"INFO: Using sheet '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")

    # Precondition: need at least 10 task rows (rows 2-11+) with data in column A
    task_count = 0
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value is not None:
            task_count += 1
    print(f"INFO: Found {task_count} task rows in column A")
    if task_count < 10:
        print(f"FAIL: Precondition — need at least 10 tasks, found {task_count}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: End Date formulas in column D (0.20 points)
    # In the initial file, column D is empty. In golden, it has formulas like =B2+C2-1
    try:
        formula_count = 0
        for r in range(2, 2 + task_count):
            d_val = ws.cell(row=r, column=4).value
            if d_val is not None and isinstance(d_val, str) and '=' in d_val:
                formula_count += 1
            elif d_val is not None and isinstance(d_val, datetime):
                # Could be a computed date (if file was saved by Calc, formula cached as value)
                formula_count += 1
        if formula_count >= 10:
            print(f"PASS: Component 1 — {formula_count} End Date formulas/values found in column D (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Expected >=10 End Date entries in column D, found {formula_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Date headers in row 1 spanning the project timeline (0.25 points)
    # Initial file has no date headers beyond column E. Golden has daily dates from col F onwards.
    try:
        date_header_count = 0
        # Check columns beyond the first 5 (A=Task, B=Start, C=Duration, D=End, E=%Complete)
        for col_idx in range(6, ws.max_column + 1):
            val = ws.cell(row=1, column=col_idx).value
            if val is not None:
                if isinstance(val, datetime):
                    date_header_count += 1
                elif isinstance(val, str):
                    try:
                        datetime.strptime(str(val), '%Y-%m-%d')
                        date_header_count += 1
                    except ValueError:
                        pass

        if date_header_count >= 20:
            # At least 20 date columns means a reasonable timeline span
            print(f"PASS: Component 2 — {date_header_count} date headers found in row 1 (0.25 pts)")
            total_score += 0.25
        elif date_header_count >= 5:
            partial = 0.25 * (date_header_count / 20.0)
            print(f"PARTIAL: Component 2 — {date_header_count} date headers (partial: {partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected date headers in row 1 beyond col E, found {date_header_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting - blue task bars (0.25 points)
    # Initial has 0 CF rules. Golden has at least one rule that creates blue bars
    # based on date range (AND formula checking if date falls within start-end range)
    try:
        cf_rules = list(ws.conditional_formatting)
        cf_count = 0
        has_blue_bar_rule = False

        for cf in cf_rules:
            for rule in cf.rules:
                cf_count += 1
                # Check for a formula-based rule with blue fill
                if rule.type == 'expression' and rule.formula:
                    formula_str = str(rule.formula[0]).upper() if rule.formula else ''
                    # Check if it's a date range comparison (AND formula with >= and <=)
                    if 'AND' in formula_str and '>=' in formula_str and '<=' in formula_str:
                        # Check for blue-ish fill color
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                fg = rule.dxf.fill.fgColor.rgb
                                if fg is not None:
                                    # Accept any blue-range color (4472C4 is standard blue)
                                    # Blue: high B component, lower R and G
                                    r_hex = int(fg[2:4], 16) if len(fg) >= 6 else 0
                                    g_hex = int(fg[4:6], 16) if len(fg) >= 6 else 0
                                    b_hex = int(fg[6:8], 16) if len(fg) >= 8 else 0
                                    if b_hex > r_hex and b_hex > 100:
                                        has_blue_bar_rule = True
                                        print(f"  INFO: Found blue bar rule with color {fg}")
                            except Exception:
                                pass

        if has_blue_bar_rule:
            print(f"PASS: Component 3 — Blue bar conditional formatting rule found ({cf_count} total CF rules) (0.25 pts)")
            total_score += 0.25
        elif cf_count > 0:
            # Has some CF rules but not clearly a blue bar rule
            print(f"PARTIAL: Component 3 — {cf_count} CF rules found but no clear blue bar rule (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting - completion color (0.15 points)
    # Golden has a second CF rule with green fill for the completed portion based on % Complete
    try:
        has_completion_rule = False
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'expression' and rule.formula:
                    formula_str = str(rule.formula[0]).upper() if rule.formula else ''
                    # The completion rule references % Complete column (E) in its formula
                    # and uses a green-ish color
                    if '$E' in formula_str or '%' in formula_str:
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                fg = rule.dxf.fill.fgColor.rgb
                                if fg is not None:
                                    # Green-ish: high G component
                                    r_hex = int(fg[2:4], 16) if len(fg) >= 6 else 0
                                    g_hex = int(fg[4:6], 16) if len(fg) >= 6 else 0
                                    b_hex = int(fg[6:8], 16) if len(fg) >= 8 else 0
                                    if g_hex > b_hex and g_hex > 100:
                                        has_completion_rule = True
                                        print(f"  INFO: Found completion rule with color {fg}")
                            except Exception:
                                pass

        if has_completion_rule:
            print(f"PASS: Component 4 — Completion percentage CF rule found (0.15 pts)")
            total_score += 0.15
        else:
            # Check if there are at least 2 different CF rules (could be completion with different approach)
            all_rules = []
            for cf in ws.conditional_formatting:
                for rule in cf.rules:
                    all_rules.append(rule)
            if len(all_rules) >= 2:
                # Two CF rules likely means one for bars and one for completion
                print(f"PARTIAL: Component 4 — {len(all_rules)} CF rules found, possibly includes completion (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 4 — No completion percentage CF rule found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Freeze panes (0.15 points)
    # Initial has no freeze. Golden has freeze_panes = D2 (freeze top row + left 3 cols A,B,C)
    try:
        freeze = ws.freeze_panes
        if freeze is not None:
            freeze_str = str(freeze)
            # The task says freeze top row and left three columns
            # D2 means rows above row 2 (row 1) and columns left of D (A,B,C) are frozen
            # Accept any freeze that freezes at least row 1 and columns A-C
            from openpyxl.utils import column_index_from_string
            col_letter = ''.join(c for c in freeze_str if c.isalpha())
            row_num = ''.join(c for c in freeze_str if c.isdigit())
            col_idx = column_index_from_string(col_letter) if col_letter else 1
            row_idx = int(row_num) if row_num else 1

            # We need col_idx >= 4 (D or beyond) to freeze A,B,C and row_idx >= 2 to freeze row 1
            if col_idx >= 4 and row_idx >= 2:
                print(f"PASS: Component 5 — Freeze panes at {freeze_str} (row 1 + cols A-C+ frozen) (0.15 pts)")
                total_score += 0.15
            elif col_idx >= 2 or row_idx >= 2:
                print(f"PARTIAL: Component 5 — Freeze panes at {freeze_str} but not fully correct (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 5 — Freeze panes at {freeze_str} does not freeze required areas")
        else:
            print(f"FAIL: Component 5 — No freeze panes set")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
