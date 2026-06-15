"""
Initial Setup: Create a project Gantt chart in LibreOffice Calc
Task ID: calc_grs_005
Domain: libreoffice_calc

Initial state: Spreadsheet with project task data (names, start dates,
durations, % complete) but NO End Date formulas, NO date headers, NO
conditional formatting, NO freeze panes, NO completion color bars.
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt Chart"

    # --- Headers in row 1 (only task-data columns, NOT date columns) ---
    headers = ["Task Name", "Start Date", "Duration (days)", "End Date", "% Complete"]
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Project task data (10 tasks for a software dev project, 8 weeks) ---
    # Project starts 2025-06-02 (a Monday), spans ~8 weeks to ~2025-07-25
    project_start = date(2025, 6, 2)

    tasks = [
        ("Requirements Gathering",       date(2025, 6, 2),   10,  15),
        ("System Architecture Design",   date(2025, 6, 9),   8,   25),
        ("Database Schema Design",       date(2025, 6, 16),  5,   40),
        ("Backend API Development",      date(2025, 6, 18),  15,  10),
        ("Frontend UI Development",      date(2025, 6, 23),  18,  5),
        ("Authentication Module",        date(2025, 6, 25),  7,   30),
        ("Integration Testing",          date(2025, 7, 7),   10,  0),
        ("Performance Optimization",     date(2025, 7, 14),  5,   0),
        ("User Acceptance Testing",      date(2025, 7, 16),  8,   0),
        ("Deployment & Documentation",   date(2025, 7, 21),  5,   0),
    ]

    data_font = Font(name="Arial", size=11)
    data_align = Alignment(horizontal="left", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    for r, (task_name, start_dt, duration, pct) in enumerate(tasks, 2):
        # Column A: Task Name
        cell_a = ws.cell(row=r, column=1, value=task_name)
        cell_a.font = data_font
        cell_a.alignment = data_align
        cell_a.border = thin_border

        # Column B: Start Date
        cell_b = ws.cell(row=r, column=2, value=start_dt)
        cell_b.font = data_font
        cell_b.alignment = center_align
        cell_b.number_format = 'yyyy-mm-dd'
        cell_b.border = thin_border

        # Column C: Duration (days)
        cell_c = ws.cell(row=r, column=3, value=duration)
        cell_c.font = data_font
        cell_c.alignment = center_align
        cell_c.border = thin_border

        # Column D: End Date - LEFT EMPTY (agent must add formula)
        cell_d = ws.cell(row=r, column=4)
        cell_d.border = thin_border

        # Column E: % Complete
        cell_e = ws.cell(row=r, column=5, value=pct / 100.0)
        cell_e.font = data_font
        cell_e.alignment = center_align
        cell_e.number_format = '0%'
        cell_e.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    # Row 1 height
    ws.row_dimensions[1].height = 25

    # NO freeze panes (agent must add)
    # NO date headers beyond column E (agent must add)
    # NO conditional formatting (agent must add)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
