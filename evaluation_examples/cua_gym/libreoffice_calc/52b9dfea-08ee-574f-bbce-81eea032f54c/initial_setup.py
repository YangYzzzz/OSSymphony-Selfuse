"""
Initial Setup: ArXiv cs.CL institution analysis spreadsheet (pre-task state)
Task ID: osworld_multi_apps_arxiv_llms_calc_014
Domain: libreoffice_calc

Creates institution_analysis.ods (as .xlsx) with:
  - Sheet1: headers only (arXiv ID, Title, First Author, Institution) — agent fills 30 rows
  - Institutions sheet: headers only (Institution, Count) — agent adds COUNTIF + chart
Opens Chrome and LibreOffice Calc so the GUI is ready.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_arxiv_llms_calc_014'
# Save as .xlsx (openpyxl native format); LibreOffice Calc opens both .xlsx and .ods
OUTPUT = f'{WORKDIR}/institution_analysis.ods'
OUTPUT_XLSX = f'{WORKDIR}/institution_analysis.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet1: papers table (headers only, no data rows yet) ----
    ws1 = wb.active
    ws1.title = 'Sheet1'

    headers1 = ['arXiv ID', 'Title', 'First Author', 'Institution']
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 20  # arXiv ID
    ws1.column_dimensions['B'].width = 55  # Title
    ws1.column_dimensions['C'].width = 25  # First Author
    ws1.column_dimensions['D'].width = 35  # Institution

    # Freeze the header row
    ws1.freeze_panes = 'A2'

    # ---- Institutions sheet: headers only (no data, no formulas, no chart) ----
    ws2 = wb.create_sheet('Institutions')

    headers2 = ['Institution', 'Count']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    ws2.column_dimensions['A'].width = 35  # Institution
    ws2.column_dimensions['B'].width = 10  # Count

    # Leave space note for chart area (just a label so agent knows where chart goes)
    ws2.cell(row=3, column=1, value='(Add institution names and COUNTIF formulas below)')
    ws2.cell(row=3, column=1).font = Font(italic=True, color='FF808080')

    # Save as .xlsx (openpyxl does not produce true ODS; xlsx is fine for LibreOffice)
    wb.save(OUTPUT_XLSX)
    import shutil
    shutil.copy(OUTPUT_XLSX, OUTPUT)
    print(f'Initial file created: {OUTPUT} (xlsx format, .ods extension)')
    print(f'Also saved as: {OUTPUT_XLSX}')

    # ---- GUI-ready startup ----
    # Launch Chrome first (agent needs to browse ArXiv)
    launch_gui('google-chrome --new-window "https://arxiv.org/list/cs.CL/2024-01"', delay_sec=3.0)

    # Launch LibreOffice Calc with the institution_analysis file
    launch_gui(f'libreoffice --calc "{OUTPUT_XLSX}"', delay_sec=3.0)

    print('GUI_READY: launched Chrome (ArXiv cs.CL Jan 2024) and LibreOffice Calc with DISPLAY=:0')


create_initial()
