"""
Reward Script: ArXiv cs.CL Institution Analysis
Task ID: osworld_multi_apps_arxiv_llms_calc_014
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): Sheet1 has 30 data rows with arXiv ID, First Author, Institution
  - Component 2 (0.35): Institutions sheet has 10 institution rows with COUNTIF formulas
  - Component 3 (0.30): Institutions sheet has a bar chart
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_arxiv_llms_calc_014'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Check Sheet1 and Institutions sheets exist
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: 'Sheet1' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    if 'Institutions' not in wb.sheetnames:
        print("CRITICAL: 'Institutions' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']
    ws2 = wb['Institutions']

    # Component 1: Sheet1 has 30 data rows (rows 2-31) with required columns (0.35 points)
    # This FAILS on initial (only 1 header row) and PASSES on golden (31 rows including header)
    try:
        data_rows = 0
        rows_with_arxiv_id = 0
        rows_with_institution = 0

        for row_idx in range(2, ws1.max_row + 1):
            arxiv_id = ws1.cell(row=row_idx, column=1).value
            institution = ws1.cell(row=row_idx, column=4).value
            if arxiv_id is not None and str(arxiv_id).strip() != '':
                data_rows += 1
                rows_with_arxiv_id += 1
            if institution is not None and str(institution).strip() != '':
                rows_with_institution += 1

        print(f"INFO: Sheet1 has {data_rows} data rows, {rows_with_arxiv_id} with arXiv IDs, {rows_with_institution} with institutions")

        if data_rows >= 30 and rows_with_institution >= 25:
            # Full credit: at least 30 rows with arXiv IDs and most have institutions
            print(f"PASS: Component 1 — Sheet1 has {data_rows} data rows with arXiv IDs and {rows_with_institution} institution values (0.35 pts)")
            total_score += 0.35
        elif data_rows >= 20 and rows_with_institution >= 15:
            # Partial credit: significant data present
            print(f"PARTIAL: Component 1 — Sheet1 has {data_rows} data rows (need 30), {rows_with_institution} institution values (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Sheet1 has {data_rows} data rows (need >= 30) and {rows_with_institution} institution values")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Institutions sheet has 10 institution rows with COUNTIF formulas (0.35 points)
    # This FAILS on initial (only placeholder text, no real data) and PASSES on golden (10 institution rows with formulas)
    try:
        institution_rows_with_data = 0
        institution_rows_with_countif = 0

        for row_idx in range(2, ws2.max_row + 1):
            inst_name = ws2.cell(row=row_idx, column=1).value
            count_formula = ws2.cell(row=row_idx, column=2).value

            if inst_name is not None and str(inst_name).strip() != '' and not str(inst_name).startswith('(Add'):
                institution_rows_with_data += 1

                # Check if the count column contains a COUNTIF formula
                if count_formula is not None:
                    formula_str = str(count_formula).upper().replace(' ', '')
                    if 'COUNTIF' in formula_str:
                        institution_rows_with_countif += 1

        print(f"INFO: Institutions sheet has {institution_rows_with_data} institution rows, {institution_rows_with_countif} with COUNTIF formulas")

        if institution_rows_with_data >= 10 and institution_rows_with_countif >= 8:
            # Full credit: at least 10 institution rows with COUNTIF formulas
            print(f"PASS: Component 2 — Institutions sheet has {institution_rows_with_data} institution rows and {institution_rows_with_countif} COUNTIF formulas (0.35 pts)")
            total_score += 0.35
        elif institution_rows_with_data >= 5 and institution_rows_with_countif >= 3:
            # Partial credit
            print(f"PARTIAL: Component 2 — Institutions sheet has {institution_rows_with_data} institution rows (need 10) and {institution_rows_with_countif} COUNTIF formulas (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Institutions sheet has {institution_rows_with_data} institution rows (need >= 10) and {institution_rows_with_countif} COUNTIF formulas (need >= 8)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Institutions sheet has a bar chart showing top 10 institutions (0.30 points)
    # This FAILS on initial (no charts) and PASSES on golden (1 BarChart)
    try:
        charts = ws2._charts
        num_charts = len(charts)
        print(f"INFO: Institutions sheet has {num_charts} chart(s)")

        if num_charts >= 1:
            # Check if any chart is a BarChart
            bar_chart_found = False
            for chart in charts:
                chart_type_name = type(chart).__name__
                if 'Bar' in chart_type_name or 'bar' in chart_type_name:
                    bar_chart_found = True
                    break
                # Also check chart type attribute
                if hasattr(chart, 'type') and chart.type in ('bar', 'col'):
                    bar_chart_found = True
                    break

            if bar_chart_found:
                print(f"PASS: Component 3 — Institutions sheet has a bar chart (0.30 pts)")
                total_score += 0.30
            else:
                # Any chart gets partial credit
                print(f"PARTIAL: Component 3 — Institutions sheet has {num_charts} chart(s) but not a bar chart (0.15 pts)")
                total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Institutions sheet has no charts (need a bar chart)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
# Task uses institution_analysis.xlsx (ods converted to xlsx by setup-gen)
file_path = f'{WORKDIR}/institution_analysis.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    # Also try the ods file
    ods_path = f'{WORKDIR}/institution_analysis.ods'
    if os.path.exists(ods_path):
        print(f"NOTE: Found .ods file but cannot verify with openpyxl, trying .xlsx")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
