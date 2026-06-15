"""
Initial Setup: Nonprofit After-School Tutoring Program Evaluation
Task ID: calc_edu_nonprofit_program_eval_068
Domain: libreoffice_calc

Creates ProgramData sheet with 8 sites, columns B-E filled (F-I empty).
The agent will calculate Score Improvement, Cost Per Student, Efficiency Score, and Rank.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_nonprofit_program_eval_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProgramData ---
    ws = wb.active
    ws.title = 'ProgramData'

    # Headers (Row 1)
    headers = [
        'Site Name',       # A
        'Students Served', # B
        'Program Cost',    # C
        'Pre-Test Avg',    # D
        'Post-Test Avg',   # E
        'Score Improvement', # F - empty (task target)
        'Cost Per Student',  # G - empty (task target)
        'Efficiency Score',  # H - empty (task target)
        'Rank',              # I - empty (task target)
    ]
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF2F75B6', end_color='FF2F75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Program site data — realistic nonprofit tutoring sites
    # Columns: Site Name, Students Served, Program Cost, Pre-Test Avg, Post-Test Avg
    # NOTE: Score Improvement, Cost Per Student, Efficiency Score, Rank are LEFT EMPTY
    site_data = [
        ('Eastside Community Center',    142, 38500.00, 58.3, 71.2),
        ('Northgate Learning Hub',        98, 29750.00, 61.7, 76.4),
        ('Riverside Youth Academy',      187, 52200.00, 55.9, 69.8),
        ('Lakewood Elementary Annex',     76, 18900.00, 63.4, 79.1),
        ('Central Park After-School',    215, 61800.00, 52.1, 66.3),
        ('West End Learning Center',     103, 27400.00, 60.2, 74.5),
        ('Hillside Neighborhood School', 134, 41600.00, 57.8, 72.6),
        ('Downtown Youth Services',       89, 24300.00, 64.1, 77.8),
    ]

    data_alignment = Alignment(horizontal='center', vertical='center')
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    currency_fmt = '$#,##0.00'

    for row_idx, (site, students, cost, pre, post) in enumerate(site_data, 2):
        # A: Site Name
        cell_a = ws.cell(row=row_idx, column=1, value=site)
        cell_a.font = Font(name='Calibri', size=11)
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        cell_a.border = data_border

        # B: Students Served
        cell_b = ws.cell(row=row_idx, column=2, value=students)
        cell_b.font = Font(name='Calibri', size=11)
        cell_b.alignment = data_alignment
        cell_b.border = data_border

        # C: Program Cost (currency format)
        cell_c = ws.cell(row=row_idx, column=3, value=cost)
        cell_c.font = Font(name='Calibri', size=11)
        cell_c.number_format = currency_fmt
        cell_c.alignment = data_alignment
        cell_c.border = data_border

        # D: Pre-Test Avg
        cell_d = ws.cell(row=row_idx, column=4, value=pre)
        cell_d.font = Font(name='Calibri', size=11)
        cell_d.number_format = '0.0'
        cell_d.alignment = data_alignment
        cell_d.border = data_border

        # E: Post-Test Avg
        cell_e = ws.cell(row=row_idx, column=5, value=post)
        cell_e.font = Font(name='Calibri', size=11)
        cell_e.number_format = '0.0'
        cell_e.alignment = data_alignment
        cell_e.border = data_border

        # F, G, H, I: EMPTY (task agent fills these in)

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 8

    # Freeze panes: lock header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: ProgramData')
    print(f'Rows: 1 header + 8 data rows')
    print(f'Columns B-E filled; F-I empty (task targets)')


create_initial()
