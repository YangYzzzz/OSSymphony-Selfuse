"""
Reward Script: Evaluate nonprofit after-school tutoring program effectiveness
Task ID: calc_edu_nonprofit_program_eval_068
Domain: libreoffice_calc
Scoring:
  - Component 1: Score Improvement formulas in F2:F9 (=En-Dn pattern)         0.25 pts
  - Component 2: Cost Per Student formulas in G2:G9 with currency format       0.25 pts
  - Component 3: Efficiency Score formulas in H2:H9 with 3-decimal format      0.25 pts
  - Component 4: Rank formulas in I2:I9 using RANK function                    0.15 pts
  - Component 5: Conditional formatting gold background on top performer row    0.10 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_nonprofit_program_eval_068'

DATA_ROWS = range(2, 10)  # rows 2-9, 8 sites


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the ProgramData sheet exists as a precondition gate
    if 'ProgramData' not in wb.sheetnames:
        print("FAIL: 'ProgramData' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ProgramData']

    # -------------------------------------------------------------------
    # Component 1: Score Improvement formulas in F2:F9 (0.25 points)
    # Each cell Fn should contain formula =En-Dn (e.g., =E2-D2)
    # Initial file has None in F cells; golden has the formula.
    # -------------------------------------------------------------------
    try:
        f_formula_count = 0
        f_expected = 8
        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=6).value  # Column F
            if cell_val is not None and isinstance(cell_val, str):
                formula = cell_val.upper().replace(' ', '')
                expected = f'=E{row}-D{row}'.upper()
                if formula == expected:
                    f_formula_count += 1
                else:
                    print(f"  WARN: F{row} formula is {repr(cell_val)}, expected =E{row}-D{row}")
            else:
                print(f"  FAIL: F{row} is empty or not a formula: {repr(cell_val)}")

        if f_formula_count == f_expected:
            print(f"PASS: Component 1 — All {f_expected} Score Improvement formulas present in F2:F9 (0.25 pts)")
            total_score += 0.25
        elif f_formula_count > 0:
            partial = round(0.25 * f_formula_count / f_expected, 4)
            print(f"PARTIAL: Component 1 — {f_formula_count}/{f_expected} Score Improvement formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No Score Improvement formulas found in F2:F9 (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: Cost Per Student formulas in G2:G9 with currency format (0.25 points)
    # Each cell Gn should contain formula =Cn/Bn, formatted as currency ($#,##0.00)
    # Initial file has None in G cells; golden has the formula + currency format.
    # -------------------------------------------------------------------
    try:
        g_formula_count = 0
        g_format_count = 0
        g_expected = 8
        for row in DATA_ROWS:
            cell = ws.cell(row=row, column=7)  # Column G
            cell_val = cell.value
            formula_matches = (
                cell_val is not None
                and isinstance(cell_val, str)
                and cell_val.upper().replace(' ', '') == f'=C{row}/B{row}'.upper()
            )
            if formula_matches:
                g_formula_count += 1
                # Check currency format
                fmt = cell.number_format
                if fmt and '$' in fmt:
                    g_format_count += 1
                else:
                    print(f"  WARN: G{row} has formula but number_format is {repr(fmt)}, expected currency format")
            else:
                print(f"  FAIL: G{row} value is {repr(cell_val)}, expected =C{row}/B{row}")

        if g_formula_count == g_expected and g_format_count == g_expected:
            print(f"PASS: Component 2 — All {g_expected} Cost Per Student formulas with currency format in G2:G9 (0.25 pts)")
            total_score += 0.25
        elif g_formula_count == g_expected:
            print(f"PARTIAL: Component 2 — All formulas present but only {g_format_count}/{g_expected} with currency format (0.15 pts)")
            total_score += 0.15
        elif g_formula_count > 0:
            partial = round(0.25 * g_formula_count / g_expected, 4)
            print(f"PARTIAL: Component 2 — {g_formula_count}/{g_expected} Cost Per Student formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No Cost Per Student formulas found in G2:G9 (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Efficiency Score formulas in H2:H9 with 3-decimal format (0.25 points)
    # Each cell Hn should contain formula =Fn/Gn*100, formatted as 0.000 (3 decimal places)
    # Initial file has None in H cells; golden has the formula + format.
    # -------------------------------------------------------------------
    try:
        h_formula_count = 0
        h_format_count = 0
        h_expected = 8
        for row in DATA_ROWS:
            cell = ws.cell(row=row, column=8)  # Column H
            cell_val = cell.value
            formula_matches = (
                cell_val is not None
                and isinstance(cell_val, str)
                and cell_val.upper().replace(' ', '') == f'=F{row}/G{row}*100'.upper()
            )
            if formula_matches:
                h_formula_count += 1
                # Check 3-decimal format (0.000)
                fmt = cell.number_format
                if fmt and '000' in fmt and '0.000' in fmt:
                    h_format_count += 1
                else:
                    print(f"  WARN: H{row} has formula but number_format is {repr(fmt)}, expected 3-decimal format (0.000)")
            else:
                print(f"  FAIL: H{row} value is {repr(cell_val)}, expected =F{row}/G{row}*100")

        if h_formula_count == h_expected and h_format_count == h_expected:
            print(f"PASS: Component 3 — All {h_expected} Efficiency Score formulas with 3-decimal format in H2:H9 (0.25 pts)")
            total_score += 0.25
        elif h_formula_count == h_expected:
            print(f"PARTIAL: Component 3 — All formulas present but only {h_format_count}/{h_expected} with 3-decimal format (0.15 pts)")
            total_score += 0.15
        elif h_formula_count > 0:
            partial = round(0.25 * h_formula_count / h_expected, 4)
            print(f"PARTIAL: Component 3 — {h_formula_count}/{h_expected} Efficiency Score formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No Efficiency Score formulas found in H2:H9 (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: Rank formulas in I2:I9 using RANK function (0.15 points)
    # Each cell In should contain formula =RANK(Hn,$H$2:$H$9,0) — rank 1 = most efficient
    # Initial file has None in I cells; golden has the formula.
    # -------------------------------------------------------------------
    try:
        i_formula_count = 0
        i_expected = 8
        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=9).value  # Column I
            formula_matches = (
                cell_val is not None
                and isinstance(cell_val, str)
                and cell_val.upper().replace(' ', '') == f'=RANK(H{row},$H$2:$H$9,0)'.upper()
            )
            if formula_matches:
                i_formula_count += 1
            else:
                print(f"  FAIL: I{row} value is {repr(cell_val)}, expected =RANK(H{row},$H$2:$H$9,0)")

        if i_formula_count == i_expected:
            print(f"PASS: Component 4 — All {i_expected} Rank formulas present in I2:I9 (0.15 pts)")
            total_score += 0.15
        elif i_formula_count > 0:
            partial = round(0.15 * i_formula_count / i_expected, 4)
            print(f"PARTIAL: Component 4 — {i_formula_count}/{i_expected} Rank formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No Rank formulas found in I2:I9 (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------
    # Component 5: Conditional formatting with gold background on top performer row (0.10 points)
    # Rule: formula $I2=1 applied to A2:I9, fill color FFFFD700 (gold #FFD700)
    # Initial file has no conditional formatting; golden has the CF rule.
    # -------------------------------------------------------------------
    try:
        cf_rule_with_gold = 0
        cf_rule_any = 0
        gold_argb = 'FFFFD700'

        for cf_range_obj in ws.conditional_formatting:
            for rule in cf_range_obj.rules:
                if rule.type == 'expression' and rule.formula:
                    formula_str = rule.formula[0].upper().replace(' ', '')
                    # Accept $I2=1 or $I$2=1 patterns
                    if 'I' in formula_str and '=1' in formula_str:
                        cf_rule_any += 1
                        # Check gold color in dxf fill
                        if rule.dxf and rule.dxf.fill:
                            try:
                                fg_rgb = rule.dxf.fill.fgColor.rgb
                                if fg_rgb and fg_rgb.upper() == gold_argb.upper():
                                    cf_rule_with_gold += 1
                                else:
                                    print(f"  WARN: CF fill color is {repr(fg_rgb)}, expected {gold_argb}")
                            except Exception as color_e:
                                print(f"  WARN: Could not read CF fill color: {color_e}")

        if cf_rule_with_gold > 0:
            print(f"PASS: Component 5 — Conditional formatting with gold background (#FFD700) on rank=1 row present (0.10 pts)")
            total_score += 0.10
        elif cf_rule_any > 0:
            print(f"PARTIAL: Component 5 — Conditional formatting rule found but gold color not confirmed (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting rule for top performer found (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------
    # Final score
    # -------------------------------------------------------------------
    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
