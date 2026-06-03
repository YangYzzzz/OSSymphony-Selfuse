"""
Initial Setup: Contact List spreadsheet with column B at default width
Task ID: calc_fmt_col_width_specific_051
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_col_width_specific_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Contact List ---
    ws = wb.active
    ws.title = 'Contact List'

    # Headers
    headers = ['ID', 'Full Name', 'Email', 'Phone', 'City']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic contact data (with long names that would be truncated)
    data = [
        [1,  'Alexandra Petrov-Smirnova',   'alexandra.ps@example.com',    '+1-555-204-8837', 'San Francisco'],
        [2,  'Christopher Worthington III',  'c.worthington3@example.com',  '+1-555-319-7642', 'Boston'],
        [3,  'Valentina Cruz-Mendoza',       'v.cruzmendoza@example.com',   '+1-555-478-1023', 'Los Angeles'],
        [4,  'Bartholomew Hutchinson',       'bart.hutchinson@example.com', '+1-555-601-9954', 'Chicago'],
        [5,  'Marguerite Delacroix-Fontaine','m.delacroix@example.com',     '+1-555-742-3315', 'New Orleans'],
        [6,  'Reginald T. Pemberton Jr.',    'reg.pemberton@example.com',   '+1-555-836-0047', 'Philadelphia'],
        [7,  'Isabella Schwarzenberger',     'i.schwarzenberger@example.com','+1-555-915-6628','Seattle'],
        [8,  'Maximilian von Bergstein',     'max.bergstein@example.com',   '+1-555-127-4439', 'Denver'],
        [9,  'Theodosia Abernathy-Cole',     't.abernathy@example.com',     '+1-555-263-8810', 'Atlanta'],
        [10, 'Cornelius Wickham-Foster',     'c.wickhamfoster@example.com', '+1-555-384-5521', 'Houston'],
        [11, 'Evangeline Duplessis',         'e.duplessis@example.com',     '+1-555-495-7732', 'Miami'],
        [12, 'Montgomery St. Claire',        'm.stclaire@example.com',      '+1-555-516-2943', 'Phoenix'],
        [13, 'Persephone Nightingale',       'p.nightingale@example.com',   '+1-555-628-0154', 'Portland'],
        [14, 'Alistair Drummond-MacLeod',    'a.drummond@example.com',      '+1-555-739-4465', 'Minneapolis'],
        [15, 'Seraphina Kowalski-Nowak',     's.kowalski@example.com',      '+1-555-840-6776', 'Detroit'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column A width (ID column — narrow)
    ws.column_dimensions['A'].width = 5
    # Column B uses default width (~8.43 characters) — DO NOT set to 25
    # Column C (Email) — reasonable width
    ws.column_dimensions['C'].width = 30
    # Column D (Phone)
    ws.column_dimensions['D'].width = 18
    # Column E (City)
    ws.column_dimensions['E'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Column B width: {ws.column_dimensions["B"].width} (default)')


create_initial()
