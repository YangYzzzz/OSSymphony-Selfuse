"""
Initial Setup: Project Timeline Tracker
Task ID: calc_wf_013
Domain: libreoffice_calc

Creates a Project Plan spreadsheet with 10 project tasks, start dates,
durations, predecessors, and status. Columns G onwards have a date grid
for Gantt visualization. End date column (D) is left empty for the agent
to fill with WORKDAY formulas. No conditional formatting applied.
"""

import os
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Project Plan'

    # --- Headers ---
    headers = ['Task', 'Start', 'Duration (days)', 'End', 'Predecessor', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # --- Project Tasks Data ---
    # Base date: 2025-09-01 (Monday)
    base_date = datetime(2025, 9, 1)

    # task_name, start_date, duration_days, predecessor (0=none), status
    tasks = [
        ('Requirements Gathering', base_date, 5, 0, 'Complete'),
        ('Stakeholder Interviews', base_date + timedelta(days=7), 3, 1, 'Complete'),
        ('System Architecture Design', base_date + timedelta(days=10), 8, 2, 'In Progress'),
        ('Database Schema Design', base_date + timedelta(days=21), 4, 3, 'In Progress'),
        ('Frontend Wireframes', base_date + timedelta(days=21), 6, 3, 'Overdue'),
        ('Backend API Development', base_date + timedelta(days=28), 15, 4, 'Overdue'),
        ('Frontend Development', base_date + timedelta(days=28), 12, 5, 'Not Started'),
        ('Integration Testing', base_date + timedelta(days=49), 5, 6, 'Not Started'),
        ('User Acceptance Testing', base_date + timedelta(days=56), 4, 8, 'Not Started'),
        ('Production Deployment', base_date + timedelta(days=63), 2, 9, 'Not Started'),
    ]

    data_font = Font(name='Calibri', size=11)
    date_format = 'yyyy-mm-dd'

    for r, (name, start, dur, pred, status) in enumerate(tasks, 2):
        # A: Task name
        cell_a = ws.cell(row=r, column=1, value=name)
        cell_a.font = data_font
        cell_a.border = thin_border

        # B: Start date
        cell_b = ws.cell(row=r, column=2, value=start)
        cell_b.font = data_font
        cell_b.number_format = date_format
        cell_b.border = thin_border
        cell_b.alignment = Alignment(horizontal='center')

        # C: Duration (days)
        cell_c = ws.cell(row=r, column=3, value=dur)
        cell_c.font = data_font
        cell_c.alignment = Alignment(horizontal='center')
        cell_c.border = thin_border

        # D: End (left EMPTY for agent to fill with WORKDAY formula)
        cell_d = ws.cell(row=r, column=4)
        cell_d.number_format = date_format
        cell_d.border = thin_border
        cell_d.alignment = Alignment(horizontal='center')

        # E: Predecessor
        if pred > 0:
            cell_e = ws.cell(row=r, column=5, value=pred)
        else:
            cell_e = ws.cell(row=r, column=5, value='None')
        cell_e.font = data_font
        cell_e.alignment = Alignment(horizontal='center')
        cell_e.border = thin_border

        # F: Status
        cell_f = ws.cell(row=r, column=6, value=status)
        cell_f.font = data_font
        cell_f.alignment = Alignment(horizontal='center')
        cell_f.border = thin_border

    # --- Date Grid for Gantt Chart (Columns G onwards) ---
    # Create a date grid spanning the project duration (~70 working days)
    # Use weekly intervals to keep the grid manageable
    gantt_start = base_date
    num_weeks = 12  # 12 weeks covers the project
    gantt_header_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    gantt_header_font = Font(name='Calibri', size=8, bold=True)

    for w in range(num_weeks):
        col = 7 + w  # G=7, H=8, ...
        date_val = gantt_start + timedelta(weeks=w)
        cell = ws.cell(row=1, column=col, value=date_val)
        cell.number_format = 'mm/dd'
        cell.font = gantt_header_font
        cell.fill = gantt_header_fill
        cell.alignment = Alignment(horizontal='center', text_rotation=90)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = 6

        # Add empty cells with borders in the data rows for the Gantt grid
        for r in range(2, 12):
            grid_cell = ws.cell(row=r, column=col)
            grid_cell.border = thin_border

    # Freeze top row and first column
    ws.freeze_panes = 'B2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
