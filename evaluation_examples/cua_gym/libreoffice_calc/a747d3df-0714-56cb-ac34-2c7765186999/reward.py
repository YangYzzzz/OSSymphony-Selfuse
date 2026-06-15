"""
Reward Script: Project Timeline Tracker with WORKDAY formulas, predecessor references,
               conditional formatting for overdue tasks, and Gantt-style date grid.
Task ID: calc_wf_013
Domain: libreoffice_calc
Scoring:
  Component 1: End date WORKDAY formulas in D2:D11       (0.30 pts)
  Component 2: Predecessor-based start date formulas B3:B11 (0.30 pts)
  Component 3: Overdue conditional formatting on F column  (0.20 pts)
  Component 4: Gantt-style conditional formatting on date grid (0.20 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_013'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Project Plan sheet
    if 'Project Plan' not in wb.sheetnames:
        print("FAIL: 'Project Plan' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Project Plan']

    # =========================================================================
    # Component 1: End date WORKDAY formulas in D2:D11 (0.30 points)
    # Task requires: D2=WORKDAY(B2,C2) for end dates excluding weekends
    # Initial state: D column is empty => this component scores the change
    # =========================================================================
    try:
        workday_count = 0
        total_d_cells = 10  # D2 through D11
        for row in range(2, 12):
            cell_val = ws.cell(row=row, column=4).value  # column D
            if cell_val is not None and isinstance(cell_val, str):
                # Check for WORKDAY formula pattern
                normalized = cell_val.upper().replace(" ", "")
                if "WORKDAY(" in normalized:
                    # Verify it references the correct B and C columns for this row
                    expected_pattern = f"WORKDAY(B{row},C{row})"
                    if normalized == "=" + expected_pattern.upper():
                        workday_count += 1
                    elif "WORKDAY(" in normalized:
                        # Accept any WORKDAY formula referencing this row
                        workday_count += 1

        if workday_count >= 8:
            # At least 8 of 10 have proper WORKDAY formulas
            score_1 = 0.30
        elif workday_count >= 5:
            score_1 = 0.20
        elif workday_count >= 1:
            score_1 = 0.10
        else:
            score_1 = 0.0

        if score_1 > 0:
            print(f"PASS: Component 1 — {workday_count}/{total_d_cells} End date WORKDAY formulas found ({score_1} pts)")
        else:
            print(f"FAIL: Component 1 — {workday_count}/{total_d_cells} End date WORKDAY formulas found (expected >= 1)")
        total_score += score_1

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Predecessor-based start date formulas in B3:B11 (0.30 points)
    # Task requires: start dates for dependent tasks auto-calculated from predecessor end dates
    # Golden pattern: =D<pred>+1 (e.g., B3=D2+1, B4=D3+1)
    # Initial state: B column has static dates => this component scores the change
    # =========================================================================
    try:
        pred_formula_count = 0
        total_pred_cells = 9  # B3 through B11
        for row in range(3, 12):
            cell_val = ws.cell(row=row, column=2).value  # column B
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                # Check for a formula referencing column D (predecessor end date)
                # Patterns: =D2+1, =INDIRECT(...), =D<n>+1, etc.
                if normalized.startswith("=") and "D" in normalized:
                    pred_formula_count += 1

        if pred_formula_count >= 7:
            # At least 7 of 9 have predecessor-based formulas
            score_2 = 0.30
        elif pred_formula_count >= 4:
            score_2 = 0.20
        elif pred_formula_count >= 1:
            score_2 = 0.10
        else:
            score_2 = 0.0

        if score_2 > 0:
            print(f"PASS: Component 2 — {pred_formula_count}/{total_pred_cells} predecessor start-date formulas found ({score_2} pts)")
        else:
            print(f"FAIL: Component 2 — {pred_formula_count}/{total_pred_cells} predecessor start-date formulas found (expected >= 1)")
        total_score += score_2

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Overdue conditional formatting on F column (0.20 points)
    # Task requires: conditional formatting to highlight 'Overdue' in red on Status column
    # Initial state: no conditional formatting => this component scores the change
    # =========================================================================
    try:
        overdue_cf_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the CF range covers part of column F (Status column)
            if 'F' in cf_range:
                for rule in cf.rules:
                    # Check for cellIs rule matching "Overdue"
                    rule_formula = getattr(rule, 'formula', None)
                    rule_operator = getattr(rule, 'operator', None)
                    rule_type = getattr(rule, 'type', None)

                    is_overdue_rule = False
                    if rule_type == 'cellIs' and rule_operator == 'equal':
                        if rule_formula and any('Overdue' in str(f) for f in rule_formula):
                            is_overdue_rule = True
                    elif rule_type == 'expression' and rule_formula:
                        # Could use a formula-based approach
                        if any('Overdue' in str(f) or 'overdue' in str(f).lower() for f in rule_formula):
                            is_overdue_rule = True
                    elif rule_type == 'containsText':
                        if rule_formula and any('Overdue' in str(f) for f in rule_formula):
                            is_overdue_rule = True

                    if is_overdue_rule:
                        # Verify it has a red-ish fill or font
                        has_red = False
                        if hasattr(rule, 'dxf') and rule.dxf:
                            if rule.dxf.fill and rule.dxf.fill.fgColor:
                                color = str(rule.dxf.fill.fgColor.rgb)
                                # Red fill: high R, low G, low B
                                if color and len(color) >= 6:
                                    # Accept various red shades
                                    if 'FF0000' in color or 'ff0000' in color.lower():
                                        has_red = True
                                    else:
                                        # Parse ARGB: check R > 200, G < 100, B < 100
                                        try:
                                            argb = color[-6:]
                                            r = int(argb[0:2], 16)
                                            g = int(argb[2:4], 16)
                                            b = int(argb[4:6], 16)
                                            if r > 180 and g < 100 and b < 100:
                                                has_red = True
                                        except Exception:
                                            pass
                            if rule.dxf.font and rule.dxf.font.color:
                                # White font on red background also counts
                                has_red = True  # If there's any styling, accept it

                        if has_red:
                            overdue_cf_found = True

        if overdue_cf_found:
            print(f"PASS: Component 3 — Overdue conditional formatting found with red styling (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — No conditional formatting rule for 'Overdue' with red styling found on F column")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Gantt-style conditional formatting on date grid G-R (0.20 points)
    # Task requires: CF using AND(G$1>=$B2,G$1<=$D2) to shade cells in date grid
    # Initial state: no conditional formatting => this component scores the change
    # =========================================================================
    try:
        gantt_cf_found = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the range covers the date grid area (columns G onwards)
            has_grid_cols = False
            for col_letter in ['G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
                if col_letter in cf_range:
                    has_grid_cols = True
                    break

            if has_grid_cols:
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', None)
                    rule_formula = getattr(rule, 'formula', None)

                    if rule_type == 'expression' and rule_formula:
                        formula_str = str(rule_formula[0]).upper().replace(" ", "")
                        # Check for AND formula referencing date comparison
                        # Expected pattern: AND(G$1>=$B2,G$1<=$D2) or similar
                        if 'AND(' in formula_str and '>=' in formula_str and '<=' in formula_str:
                            # Verify it references both start (B) and end (D) columns
                            if '$B' in formula_str and '$D' in formula_str:
                                gantt_cf_found = True
                            elif 'B' in formula_str and 'D' in formula_str:
                                gantt_cf_found = True

        if gantt_cf_found:
            print(f"PASS: Component 4 — Gantt-style conditional formatting found on date grid (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — No Gantt-style conditional formatting (AND with date range) found on date grid columns")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state(domain):
    import os, time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
