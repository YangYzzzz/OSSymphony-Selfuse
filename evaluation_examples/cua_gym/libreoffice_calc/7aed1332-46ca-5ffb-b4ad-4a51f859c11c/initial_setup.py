"""
Initial Setup: Apply custom number format for +/- percentages
Task ID: calc_gfl_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Variances'

    # Headers
    headers = ['Metric', 'Variance', 'Target', 'Actual', 'Note']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 19 rows of realistic variance data (rows 2-20)
    data = [
        ['Revenue Growth', 0.0525, 1500000, 1578750, 'Strong Q1 performance'],
        ['Cost of Goods Sold', -0.031, 820000, 845420, 'Raw material price increase'],
        ['Operating Margin', 0.0182, 0.22, 0.2382, 'Efficiency gains in production'],
        ['Customer Acquisition', -0.0475, 2500, 2618.75, 'Higher ad spend than planned'],
        ['Employee Retention', 0.0089, 0.92, 0.9289, 'New benefits program effective'],
        ['Inventory Turnover', -0.0156, 8.5, 8.6326, 'Seasonal slowdown in sales'],
        ['Net Profit Margin', 0.0340, 0.15, 0.1840, 'Tax optimization strategy'],
        ['Market Share', 0.0215, 0.12, 0.1415, 'Competitor exit from segment'],
        ['R&D Spending', -0.0623, 450000, 478035, 'Additional prototype costs'],
        ['Customer Satisfaction', 0.0147, 4.2, 4.2617, 'Service desk improvements'],
        ['Debt-to-Equity Ratio', -0.0083, 0.65, 0.6554, 'Scheduled loan repayment'],
        ['Working Capital', 0.0291, 380000, 391058, 'Faster receivables collection'],
        ['Return on Assets', -0.0198, 0.08, 0.0816, 'Asset depreciation adjustment'],
        ['Sales per Employee', 0.0412, 125000, 130150, 'Headcount optimization'],
        ['Marketing ROI', -0.0567, 3.8, 4.0155, 'Campaign underperformance'],
        ['Supply Chain Cost', 0.0133, 620000, 611754, 'Renegotiated vendor contracts'],
        ['Quality Defect Rate', -0.0024, 0.015, 0.01504, 'Minor calibration issue'],
        ['Energy Consumption', 0.0376, 28500, 27428.60, 'LED retrofit savings'],
        ['Training Budget Util.', -0.0251, 0.85, 0.8713, 'Delayed onboarding sessions'],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Metric
        ws.cell(row=r, column=2, value=row_data[1])  # Variance (decimal)
        ws.cell(row=r, column=3, value=row_data[2])  # Target
        ws.cell(row=r, column=4, value=row_data[3])  # Actual
        ws.cell(row=r, column=5, value=row_data[4])  # Note

    # Set column widths for readability
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 34

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
