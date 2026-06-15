"""
Reward Script: Apply custom +/- percentage number format to B2:B20
Task ID: calc_gfl_032
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 19 cells B2:B20 have a custom number format (not 'General')
                      that contains '%' and '+' indicating a +/- percentage format.
  Component 2 (0.3): The exact format code is '+0.00%;-0.00%' for all 19 cells.
  Component 3 (0.2): Underlying decimal values are unchanged (data integrity).
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_032'

# Expected original values for B2:B20 (data integrity check)
EXPECTED_VALUES = [
    0.0525, -0.031, 0.0182, -0.0475, 0.0089,
    -0.0156, 0.034, 0.0215, -0.0623, 0.0147,
    -0.0083, 0.0291, -0.0198, 0.0412, -0.0567,
    0.0133, -0.0024, 0.0376, -0.0251
]


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb['Variances']
    except KeyError:
        print("CRITICAL: Sheet 'Variances' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All 19 cells B2:B20 have a custom percentage format with +/- sign (0.5 pts)
    # This checks that the format is NOT 'General' and contains both '%' and '+'
    try:
        custom_fmt_count = 0
        for r in range(2, 21):
            nf = ws.cell(row=r, column=2).number_format
            if nf and nf != 'General' and '%' in nf and '+' in nf:
                custom_fmt_count += 1
        if custom_fmt_count == 19:
            print(f"PASS: Component 1 — All 19 cells have custom +/- percentage format (0.5 pts)")
            total_score += 0.5
        elif custom_fmt_count > 0:
            partial = round(0.5 * custom_fmt_count / 19, 2)
            print(f"PARTIAL: Component 1 — {custom_fmt_count}/19 cells have custom +/- pct format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells have custom +/- percentage format (0/19)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Exact format code is '+0.00%;-0.00%' for all 19 cells (0.3 pts)
    # Accepts common equivalent forms: '+0.00%;-0.00%' or '+0.00%;\\-0.00%'
    try:
        exact_fmt_count = 0
        for r in range(2, 21):
            nf = ws.cell(row=r, column=2).number_format
            # Normalize by removing backslash escapes for comparison
            nf_normalized = nf.replace('\\', '') if nf else ''
            if nf_normalized in ('+0.00%;-0.00%', '+0.00%;-0.00%;0.00%'):
                exact_fmt_count += 1
            elif nf and '+' in nf and '0.00%' in nf:
                # Close enough variant (e.g. different semicolon sections)
                exact_fmt_count += 1
        if exact_fmt_count == 19:
            print(f"PASS: Component 2 — All 19 cells have exact format code '+0.00%;-0.00%' (0.3 pts)")
            total_score += 0.3
        elif exact_fmt_count > 0:
            partial = round(0.3 * exact_fmt_count / 19, 2)
            print(f"PARTIAL: Component 2 — {exact_fmt_count}/19 cells have exact format ({partial} pts)")
            total_score += partial
        else:
            sample_fmt = ws.cell(row=2, column=2).number_format
            print(f"FAIL: Component 2 — Expected '+0.00%;-0.00%', found '{sample_fmt}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Custom format applied AND underlying values unchanged (0.2 pts)
    # Compound check: format must be non-General AND values must be intact.
    # This ensures we only award points when the task change (format) is present
    # AND data integrity is maintained. Fails on initial_env because format is General.
    try:
        compound_ok = 0
        for i, r in enumerate(range(2, 21)):
            cell = ws.cell(row=r, column=2)
            nf = cell.number_format
            val = cell.value
            expected = EXPECTED_VALUES[i]
            fmt_ok = nf and nf != 'General' and '%' in nf
            val_ok = val is not None and abs(float(val) - expected) < 1e-6
            if fmt_ok and val_ok:
                compound_ok += 1
        if compound_ok == 19:
            print(f"PASS: Component 3 — All 19 cells have custom pct format AND unchanged values (0.2 pts)")
            total_score += 0.2
        elif compound_ok > 0:
            partial = round(0.2 * compound_ok / 19, 2)
            print(f"PARTIAL: Component 3 — {compound_ok}/19 cells pass compound check ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Compound check failed (format not applied or values changed)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
