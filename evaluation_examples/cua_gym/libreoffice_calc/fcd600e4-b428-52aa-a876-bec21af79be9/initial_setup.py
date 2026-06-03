"""
Initial Setup: Law school admission trends - creates blank pass rate table and PDF reports
Task ID: osworld_multi_apps_ecs_multi_report_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

import openpyxl
from fpdf import FPDF

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_multi_apps_ecs_multi_report_010'
OUTPUT = f'{WORKDIR}/law_admissions.xlsx'
PDF_DIR = f'{WORKDIR}/Documents/LawAdmissions'

# Realistic admission rate data (%) from PDF reports
# These are the values the agent must read from PDFs and enter into the spreadsheet
ADMISSION_DATA = {
    2020: {
        'Harvard':  12.9,
        'Yale':      6.9,
        'Columbia': 17.5,
        'Chicago':  17.8,
        'NYU':      24.6,
    },
    2021: {
        'Harvard':  13.1,
        'Yale':      6.6,
        'Columbia': 16.8,
        'Chicago':  18.3,
        'NYU':      23.9,
    },
    2022: {
        'Harvard':  12.5,
        'Yale':      6.2,
        'Columbia': 15.7,
        'Chicago':  17.2,
        'NYU':      22.8,
    },
    2023: {
        'Harvard':  11.8,
        'Yale':      5.9,
        'Columbia': 14.9,
        'Chicago':  16.5,
        'NYU':      22.1,
    },
}

SCHOOLS = ['Harvard', 'Yale', 'Columbia', 'Chicago', 'NYU']
YEARS = [2020, 2021, 2022, 2023]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_pdf_report(year: int, data: dict, output_path: str):
    """Create a realistic law school admission report PDF for the given year."""
    pdf = FPDF()
    pdf.add_page()
    pdf.set_auto_page_break(auto=True, margin=15)

    # Title
    pdf.set_font("Helvetica", "B", 18)
    pdf.cell(0, 12, f"Law School Admissions Annual Report {year}", ln=True, align="C")
    pdf.ln(4)

    # Subtitle
    pdf.set_font("Helvetica", "I", 12)
    pdf.cell(0, 8, "Admission Statistics for Top-Ranked Law Schools", ln=True, align="C")
    pdf.ln(8)

    # Introduction paragraph
    pdf.set_font("Helvetica", "", 11)
    intro = (
        f"This report summarizes admission statistics for the {year} academic year "
        "at five leading law schools in the United States. The data below reflects "
        "the percentage of applicants who received offers of admission. "
        "Admission rates vary based on application volume, class size targets, "
        "and institutional selectivity goals."
    )
    pdf.multi_cell(0, 6, intro)
    pdf.ln(6)

    # Section header
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 8, f"Admission Pass Rates - {year}", ln=True)
    pdf.ln(3)

    # Table header
    pdf.set_fill_color(40, 60, 120)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(100, 9, "Law School", border=1, fill=True)
    pdf.cell(60, 9, "Admission Rate (%)", border=1, fill=True, ln=True)

    # Table rows
    pdf.set_text_color(0, 0, 0)
    row_colors = [(245, 245, 252), (255, 255, 255)]
    for i, school in enumerate(SCHOOLS):
        rate = data[school]
        pdf.set_fill_color(*row_colors[i % 2])
        pdf.set_font("Helvetica", "", 11)
        pdf.cell(100, 9, f"  {school} Law School", border=1, fill=True)
        pdf.cell(60, 9, f"  {rate:.1f}%", border=1, fill=True, ln=True)

    pdf.ln(8)

    # Notes section
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 7, "Methodology Notes:", ln=True)
    pdf.set_font("Helvetica", "", 10)
    notes = [
        "- Admission rate is calculated as (admitted applicants / total applicants) x 100.",
        "- Data is sourced from official institutional Common Data Sets (CDS).",
        "- Figures reflect first-year JD program admissions only.",
        "- Rates may differ slightly from those reported by US News & World Report.",
    ]
    for note in notes:
        pdf.cell(0, 6, note, ln=True)

    pdf.ln(6)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(
        0, 6,
        f"Published by the Law School Admissions Council (LSAC) - {year} Annual Summary",
        ln=True, align="C"
    )

    pdf.output(output_path)
    print(f"  Created PDF: {output_path}")


def create_initial():
    # ----------------------------------------------------------------
    # 1. Create PDF reports in ~/Documents/LawAdmissions
    # ----------------------------------------------------------------
    os.makedirs(PDF_DIR, exist_ok=True)
    for year in YEARS:
        pdf_path = os.path.join(PDF_DIR, f"law_admissions_{year}.pdf")
        create_pdf_report(year, ADMISSION_DATA[year], pdf_path)
    print(f"PDF reports created in: {PDF_DIR}")

    # ----------------------------------------------------------------
    # 2. Create the blank law_admissions.xlsx with stub table
    # ----------------------------------------------------------------
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Admission Rates"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # --- Header row: school label + year columns ---
    header_font = Font(name="Calibri", bold=True, size=12)
    header_fill = PatternFill(start_color="FF1F3864", end_color="FF1F3864", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(name="Calibri", bold=True, size=12, color="FFFFFFFF")

    thin = Side(style="thin", color="FF000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Row 1: headers
    ws.cell(row=1, column=1, value="Law School").font = white_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=1).border = border

    for col_idx, year in enumerate(YEARS, start=2):
        cell = ws.cell(row=1, column=col_idx, value=year)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Rows 2-6: school names, blank data cells
    school_fill_even = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    school_fill_odd  = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    school_font = Font(name="Calibri", bold=True, size=11)
    data_font   = Font(name="Calibri", size=11)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    for row_idx, school in enumerate(SCHOOLS, start=2):
        fill = school_fill_even if row_idx % 2 == 0 else school_fill_odd

        # School name cell
        name_cell = ws.cell(row=row_idx, column=1, value=school)
        name_cell.font = school_font
        name_cell.fill = fill
        name_cell.alignment = left_align
        name_cell.border = border

        # Data cells — intentionally BLANK (agent fills from PDFs)
        for col_idx in range(2, len(YEARS) + 2):
            data_cell = ws.cell(row=row_idx, column=col_idx, value=None)
            data_cell.font = data_font
            data_cell.fill = fill
            data_cell.alignment = center_align
            data_cell.border = border

    # Column widths
    ws.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E"]:
        ws.column_dimensions[col_letter].width = 12

    # Row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, 7):
        ws.row_dimensions[r].height = 20

    wb.save(OUTPUT)
    print(f"Blank spreadsheet created: {OUTPUT}")

    # ----------------------------------------------------------------
    # 3. GUI-ready startup: open spreadsheet in Calc + Nautilus
    # ----------------------------------------------------------------
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui(f'nautilus "{PDF_DIR}"', delay_sec=1.5)
    print("GUI_READY: launched LibreOffice Calc and Nautilus with DISPLAY=:0")


create_initial()
