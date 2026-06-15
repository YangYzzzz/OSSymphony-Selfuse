"""
Reward Script: Fill law school admission rate table from PDF reports
Task ID: osworld_multi_apps_ecs_multi_report_010
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): All 20 data cells (B2:E6) are filled with numeric values
  - Component 2 (0.3): All values are valid admission rates (between 0 and 100)
  - Component 3 (0.3): Specific values match expected golden values within tolerance
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_ecs_multi_report_010'

# Expected golden values: {school_row: {year_col: expected_rate}}
# Rows: Harvard=2, Yale=3, Columbia=4, Chicago=5, NYU=6
# Columns: 2020=B=2, 2021=C=3, 2022=D=4, 2023=E=5
EXPECTED_VALUES = {
    2: {2: 12.9, 3: 13.1, 4: 12.5, 5: 11.8},  # Harvard
    3: {2: 6.9,  3: 6.6,  4: 6.2,  5: 5.9},   # Yale
    4: {2: 17.5, 3: 16.8, 4: 15.7, 5: 14.9},  # Columbia
    5: {2: 17.8, 3: 18.3, 4: 17.2, 5: 16.5},  # Chicago
    6: {2: 24.6, 3: 23.9, 4: 22.8, 5: 22.1},  # NYU
}

SCHOOL_NAMES = {2: 'Harvard', 3: 'Yale', 4: 'Columbia', 5: 'Chicago', 6: 'NYU'}
YEAR_NAMES = {2: '2020', 3: '2021', 4: '2022', 5: '2023'}
TOLERANCE = 0.5  # Allow 0.5% tolerance for rounding/transcription differences


def verify_task(file_path):
    """
    Verify that the law school admission rate table has been completely filled.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify the expected sheet and structure exist
    try:
        ws = wb.active
        if ws is None:
            print("CRITICAL: No active sheet found")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition check: ensure headers are correct (gate, not scored)
    try:
        header_ok = (
            ws.cell(row=1, column=1).value == 'Law School' and
            ws.cell(row=2, column=1).value == 'Harvard' and
            ws.cell(row=3, column=1).value == 'Yale'
        )
        if not header_ok:
            print("CRITICAL: Table headers/structure are corrupted")
            print("REWARD: 0.0")
            return 0.0
        print("PRE-CHECK: Table structure intact (headers present)")
    except Exception as e:
        print(f"CRITICAL: Cannot verify headers: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All 20 data cells (B2:E6) are filled with non-null numeric values (0.4 points)
    # Rows 2-6 (5 schools), Columns 2-5 (years 2020-2023) = 20 cells total
    try:
        filled_count = 0
        missing_cells = []
        non_numeric_cells = []

        for row in range(2, 7):   # rows 2-6 (Harvard through NYU)
            for col in range(2, 6):  # cols 2-5 (2020-2023)
                cell_val = ws.cell(row=row, column=col).value
                school = SCHOOL_NAMES[row]
                year = YEAR_NAMES[col]
                if cell_val is None:
                    missing_cells.append(f"{school}/{year}")
                else:
                    try:
                        float(cell_val)
                        filled_count += 1
                    except (ValueError, TypeError):
                        non_numeric_cells.append(f"{school}/{year}={repr(cell_val)}")

        if filled_count == 20:
            print(f"PASS: Component 1 — All 20 data cells filled with numeric values (0.4 pts)")
            total_score += 0.4
        elif filled_count >= 15:
            partial = round(0.4 * filled_count / 20, 3)
            print(f"PARTIAL: Component 1 — {filled_count}/20 cells filled ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {filled_count}/20 cells filled")
            if missing_cells:
                print(f"  Missing: {', '.join(missing_cells[:5])}" + ("..." if len(missing_cells) > 5 else ""))
            if non_numeric_cells:
                print(f"  Non-numeric: {', '.join(non_numeric_cells)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All values are valid admission rates (between 0 and 100) (0.3 points)
    # This ensures values are actual percentage rates, not corrupted data
    try:
        valid_range_count = 0
        invalid_cells = []
        for row in range(2, 7):
            for col in range(2, 6):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val is not None:
                    try:
                        rate = float(cell_val)
                        if 0.0 < rate < 100.0:
                            valid_range_count += 1
                        else:
                            school = SCHOOL_NAMES[row]
                            year = YEAR_NAMES[col]
                            invalid_cells.append(f"{school}/{year}={rate}")
                    except (ValueError, TypeError):
                        pass  # Already caught in Component 1

        if valid_range_count == 20:
            print(f"PASS: Component 2 — All 20 values are valid admission rates in (0,100) range (0.3 pts)")
            total_score += 0.3
        elif valid_range_count >= 15:
            partial = round(0.3 * valid_range_count / 20, 3)
            print(f"PARTIAL: Component 2 — {valid_range_count}/20 values in valid range ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {valid_range_count}/20 values in valid range (0,100)")
            if invalid_cells:
                print(f"  Out-of-range: {', '.join(invalid_cells)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Values match expected golden values within tolerance (0.3 points)
    # Spot-check all 20 cells against known expected values derived from task requirements
    try:
        match_count = 0
        mismatch_cells = []
        total_checks = 0

        for row in range(2, 7):
            for col in range(2, 6):
                cell_val = ws.cell(row=row, column=col).value
                if cell_val is None:
                    total_checks += 1
                    school = SCHOOL_NAMES[row]
                    year = YEAR_NAMES[col]
                    mismatch_cells.append(f"{school}/{year}: found None, expected ~{EXPECTED_VALUES[row][col]}")
                    continue
                try:
                    actual = float(cell_val)
                    expected = EXPECTED_VALUES[row][col]
                    total_checks += 1
                    if abs(actual - expected) <= TOLERANCE:
                        match_count += 1
                    else:
                        school = SCHOOL_NAMES[row]
                        year = YEAR_NAMES[col]
                        mismatch_cells.append(f"{school}/{year}: found {actual}, expected {expected}")
                except (ValueError, TypeError):
                    total_checks += 1

        if match_count == 20:
            print(f"PASS: Component 3 — All 20 values match expected golden values (0.3 pts)")
            total_score += 0.3
        elif match_count >= 15:
            partial = round(0.3 * match_count / 20, 3)
            print(f"PARTIAL: Component 3 — {match_count}/20 values match expected ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {match_count}/20 values match expected golden values")
            if mismatch_cells:
                for m in mismatch_cells[:5]:
                    print(f"  {m}")
                if len(mismatch_cells) > 5:
                    print(f"  ...and {len(mismatch_cells)-5} more")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 3)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/law_admissions.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
