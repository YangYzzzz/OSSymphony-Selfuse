"""
Reward Script: Build a correlation analysis section with scatter plot and formatted coefficient matrix.
Task ID: calc_gpm_044
Domain: libreoffice_calc
Scoring:
  Component 1 — CORREL formulas in G4:I6 (0.30)
  Component 2 — Number format 0.000 on correlation cells (0.10)
  Component 3 — Diagonal cells (G4, H5, I6) have gray fill (0.10)
  Component 4 — Conditional formatting rules on G4:I6 (0.15)
  Component 5 — Scatter chart 1: 'Study Hours vs Test Score' with trendline (0.20)
  Component 6 — Scatter chart 2: 'Age vs Score' with trendline (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_044'


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def get_chart_title_text(chart):
    """Extract the plain-text title string from an openpyxl chart."""
    try:
        if chart.title is None:
            return None
        title_obj = chart.title
        # title.tx.rich.paragraphs[0].runs[0].text
        if hasattr(title_obj, 'tx') and title_obj.tx is not None:
            rich = getattr(title_obj.tx, 'rich', None)
            if rich is not None:
                for para in rich.paragraphs:
                    for run in para.r:
                        if run.t:
                            return run.t
        # Fallback: title might be a plain string
        if hasattr(title_obj, 'text'):
            return title_obj.text
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Correlation' sheet must exist
    if 'Correlation' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Correlation' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Correlation']

    # ========================================================================
    # Component 1: CORREL formulas in G4:I6 (0.30 points)
    # All 9 cells of the 3x3 matrix should contain CORREL formulas
    # ========================================================================
    try:
        correl_cells = [
            'G4', 'H4', 'I4',
            'G5', 'H5', 'I5',
            'G6', 'H6', 'I6',
        ]
        correl_count = 0
        for coord in correl_cells:
            val = ws[coord].value
            if val is not None and isinstance(val, str) and 'CORREL' in val.upper():
                correl_count += 1
            else:
                print(f"  DETAIL: {coord} has value '{val}' — no CORREL formula")

        if correl_count == 9:
            print(f"PASS: Component 1 — All 9 correlation cells have CORREL formulas (0.30 pts)")
            total_score += 0.30
        elif correl_count >= 6:
            partial = round(0.30 * (correl_count / 9), 2)
            print(f"PARTIAL: Component 1 — {correl_count}/9 CORREL formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correl_count}/9 CORREL formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ========================================================================
    # Component 2: Number format 0.000 on correlation cells (0.10 points)
    # ========================================================================
    try:
        fmt_count = 0
        for coord in correl_cells:
            nf = ws[coord].number_format
            if nf == '0.000':
                fmt_count += 1

        if fmt_count == 9:
            print(f"PASS: Component 2 — All 9 correlation cells have number format '0.000' (0.10 pts)")
            total_score += 0.10
        elif fmt_count >= 5:
            partial = round(0.10 * (fmt_count / 9), 2)
            print(f"PARTIAL: Component 2 — {fmt_count}/9 cells have '0.000' format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {fmt_count}/9 cells have '0.000' format")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ========================================================================
    # Component 3: Diagonal cells (G4, H5, I6) have gray fill (0.10 points)
    # Self-correlation cells should have a distinct gray background
    # ========================================================================
    try:
        diag_cells = ['G4', 'H5', 'I6']
        diag_gray_count = 0
        for coord in diag_cells:
            cell = ws[coord]
            fill_type = cell.fill.fill_type
            fg_rgb = getattr(cell.fill.fgColor, 'rgb', None) if cell.fill.fgColor else None
            # Accept any gray-ish solid fill (C0C0C0, 808080, A9A9A9, D3D3D3, BFBFBF, etc.)
            if fill_type == 'solid' and fg_rgb is not None:
                # Extract RGB portion (last 6 chars of ARGB)
                rgb_hex = fg_rgb[-6:] if len(fg_rgb) >= 6 else fg_rgb
                r_val = int(rgb_hex[0:2], 16)
                g_val = int(rgb_hex[2:4], 16)
                b_val = int(rgb_hex[4:6], 16)
                # Gray: all channels roughly equal and in mid-range
                if (abs(r_val - g_val) < 30 and abs(g_val - b_val) < 30
                        and r_val > 100 and r_val < 240):
                    diag_gray_count += 1
                    print(f"  DETAIL: {coord} has gray fill {fg_rgb}")
                else:
                    print(f"  DETAIL: {coord} has fill {fg_rgb} — not recognized as gray")
            else:
                print(f"  DETAIL: {coord} fill_type={fill_type}, fgColor={fg_rgb}")

        if diag_gray_count == 3:
            print(f"PASS: Component 3 — All 3 diagonal cells have gray fill (0.10 pts)")
            total_score += 0.10
        elif diag_gray_count >= 1:
            partial = round(0.10 * (diag_gray_count / 3), 2)
            print(f"PARTIAL: Component 3 — {diag_gray_count}/3 diagonal cells have gray fill ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No diagonal cells have gray fill")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ========================================================================
    # Component 4: Conditional formatting on G4:I6 (0.15 points)
    # Should have rules for positive (>0.5), negative (<-0.5), weak correlation
    # ========================================================================
    try:
        cf_rules_found = 0
        has_positive_rule = False
        has_negative_rule = False
        has_weak_rule = False

        for cf in ws.conditional_formatting:
            cf_range_str = str(cf)
            # Check if this CF applies to the correlation matrix area
            # Could be G4:I6 or similar
            if any(c in cf_range_str for c in ['G4', 'G5', 'G6', 'H4', 'H5', 'H6', 'I4', 'I5', 'I6']):
                for rule in cf.rules:
                    formulas = getattr(rule, 'formula', []) or []
                    operator = getattr(rule, 'operator', None)
                    if operator == 'greaterThan' and any('0.5' in str(f) for f in formulas):
                        has_positive_rule = True
                    elif operator == 'lessThan' and any('-0.5' in str(f) or '0.5' in str(f) for f in formulas):
                        has_negative_rule = True
                    elif operator == 'between':
                        has_weak_rule = True
                    cf_rules_found += 1

        rules_matched = sum([has_positive_rule, has_negative_rule, has_weak_rule])
        if rules_matched >= 3:
            print(f"PASS: Component 4 — Conditional formatting with positive/negative/weak rules (0.15 pts)")
            total_score += 0.15
        elif rules_matched >= 1:
            partial = round(0.15 * (rules_matched / 3), 2)
            print(f"PARTIAL: Component 4 — {rules_matched}/3 CF rule types found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found on correlation matrix")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ========================================================================
    # Component 5: Scatter chart 1 — 'Study Hours vs Test Score' (0.20 points)
    # Must be a scatter chart with correct title and linear trendline with R-squared
    # ========================================================================
    try:
        charts = ws._charts
        chart1_found = False
        chart1_trendline = False
        chart1_rsquared = False

        for chart in charts:
            title_text = get_chart_title_text(chart)
            if title_text and 'hours' in title_text.lower() and 'score' in title_text.lower():
                chart1_found = True
                # Check for scatter type
                type_name = type(chart).__name__
                print(f"  DETAIL: Chart '{title_text}' is type {type_name}")
                # Check trendline
                for series in chart.series:
                    tl = getattr(series, 'trendline', None)
                    if tl is not None:
                        chart1_trendline = True
                        tl_type = getattr(tl, 'trendlineType', None)
                        disp_rsqr = getattr(tl, 'dispRSqr', False)
                        print(f"  DETAIL: Trendline type={tl_type}, dispRSqr={disp_rsqr}")
                        if disp_rsqr:
                            chart1_rsquared = True
                break

        sub_score = 0.0
        if chart1_found:
            sub_score += 0.10  # chart exists with correct title
        if chart1_trendline:
            sub_score += 0.05  # has trendline
        if chart1_rsquared:
            sub_score += 0.05  # R-squared displayed

        if sub_score > 0:
            print(f"PASS: Component 5 — Scatter chart 'Hours vs Score' (found={chart1_found}, trendline={chart1_trendline}, R2={chart1_rsquared}) ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 5 — No scatter chart with 'Hours vs Score' title found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ========================================================================
    # Component 6: Scatter chart 2 — 'Age vs Score' (0.15 points)
    # Must be a scatter chart with correct title and linear trendline
    # ========================================================================
    try:
        chart2_found = False
        chart2_trendline = False
        chart2_rsquared = False

        for chart in charts:
            title_text = get_chart_title_text(chart)
            if title_text and 'age' in title_text.lower() and 'score' in title_text.lower():
                chart2_found = True
                type_name = type(chart).__name__
                print(f"  DETAIL: Chart '{title_text}' is type {type_name}")
                for series in chart.series:
                    tl = getattr(series, 'trendline', None)
                    if tl is not None:
                        chart2_trendline = True
                        disp_rsqr = getattr(tl, 'dispRSqr', False)
                        if disp_rsqr:
                            chart2_rsquared = True
                break

        sub_score = 0.0
        if chart2_found:
            sub_score += 0.08  # chart exists with correct title
        if chart2_trendline:
            sub_score += 0.04  # has trendline
        if chart2_rsquared:
            sub_score += 0.03  # R-squared displayed

        if sub_score > 0:
            print(f"PASS: Component 6 — Scatter chart 'Age vs Score' (found={chart2_found}, trendline={chart2_trendline}, R2={chart2_rsquared}) ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 6 — No scatter chart with 'Age vs Score' title found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state (LibreOffice may have unsaved changes)
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
