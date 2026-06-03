"""
Initial Setup: Build a correlation analysis section with scatter plot and formatted coefficient matrix.
Task ID: calc_gpm_044
Domain: libreoffice_calc

Creates: /home/user/calc_gpm_044.xlsx with raw data and matrix labels (no formulas, no charts, no conditional formatting).
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Correlation"

    # --- Merged title row A1:H1 ---
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Variable Correlation Analysis"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FF2F4F4F", end_color="FF2F4F4F", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Data section headers A3:D3 ---
    headers = ["Subject", "Age", "Score", "Hours Studied"]
    header_fill = PatternFill(start_color="FF2F4F4F", end_color="FF2F4F4F", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # --- 20 subjects with realistic data ---
    subjects = [
        ("S001", 21, 82, 8),
        ("S002", 19, 67, 4),
        ("S003", 22, 91, 12),
        ("S004", 18, 58, 3),
        ("S005", 20, 75, 7),
        ("S006", 23, 88, 11),
        ("S007", 19, 62, 5),
        ("S008", 24, 95, 14),
        ("S009", 20, 71, 6),
        ("S010", 21, 84, 9),
        ("S011", 18, 55, 2),
        ("S012", 22, 79, 8),
        ("S013", 25, 93, 13),
        ("S014", 19, 64, 4),
        ("S015", 23, 87, 10),
        ("S016", 20, 73, 6),
        ("S017", 21, 90, 12),
        ("S018", 24, 98, 15),
        ("S019", 18, 60, 3),
        ("S020", 22, 76, 7),
    ]

    for r, (subj, age, score, hours) in enumerate(subjects, 4):
        ws.cell(row=r, column=1, value=subj)
        ws.cell(row=r, column=2, value=age)
        ws.cell(row=r, column=3, value=score)
        ws.cell(row=r, column=4, value=hours)

    # --- Correlation matrix labels (F3:I6) ---
    # Column headers
    ws.cell(row=3, column=6, value="")   # F3 empty
    ws.cell(row=3, column=7, value="Age")
    ws.cell(row=3, column=8, value="Score")
    ws.cell(row=3, column=9, value="Hours")

    # Row headers
    ws.cell(row=4, column=6, value="Age")
    ws.cell(row=5, column=6, value="Score")
    ws.cell(row=6, column=6, value="Hours")

    # Bold the matrix headers
    for r in range(3, 7):
        for c in range(6, 10):
            cell = ws.cell(row=r, column=c)
            if cell.value:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 10

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
