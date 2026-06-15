"""
Initial Setup: Merge and format header cell in Summary sheet
Task ID: calc_ggf_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = 'Summary'

    # Title in A1 only - NO merge, NO special formatting (default style)
    ws['A1'] = 'Annual Financial Summary'

    # Headers in row 2
    headers = ['Category', 'Q1 ($)', 'Q2 ($)', 'Q3 ($)', 'Q4 ($)']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Financial data rows 3-14
    data = [
        ['Revenue - Product Sales', 245000, 267500, 312000, 298000],
        ['Revenue - Services', 89000, 95200, 102400, 110800],
        ['Revenue - Licensing', 34500, 36200, 38100, 41000],
        ['Cost of Goods Sold', -122500, -133750, -156000, -149000],
        ['Gross Profit', 246000, 265150, 296500, 300800],
        ['Operating Expenses', -98400, -101200, -105600, -108900],
        ['Marketing & Advertising', -45000, -52000, -48000, -55000],
        ['Research & Development', -67000, -69500, -72000, -74500],
        ['General & Administrative', -31200, -32400, -33800, -35100],
        ['Net Operating Income', 4400, 10050, 37100, 27300],
        ['Interest Income', 2100, 2300, 2500, 2700],
        ['Net Income Before Tax', 6500, 12350, 39600, 30000],
    ]
    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # --- Sheet 2: Q1 Details ---
    ws2 = wb.create_sheet('Q1 Details')
    ws2['A1'] = 'Q1 Detailed Breakdown'
    q1_headers = ['Month', 'Product Sales', 'Services', 'Licensing', 'Total']
    for col, h in enumerate(q1_headers, 1):
        cell = ws2.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
    q1_data = [
        ['January', 78000, 28500, 11000, 117500],
        ['February', 82000, 29800, 11500, 123300],
        ['March', 85000, 30700, 12000, 127700],
    ]
    for r, row_data in enumerate(q1_data, 3):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Sheet 3: Q2 Details ---
    ws3 = wb.create_sheet('Q2 Details')
    ws3['A1'] = 'Q2 Detailed Breakdown'
    for col, h in enumerate(q1_headers, 1):
        cell = ws3.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True)
    q2_data = [
        ['April', 86000, 31000, 11800, 128800],
        ['May', 89500, 32200, 12100, 133800],
        ['June', 92000, 32000, 12300, 136300],
    ]
    for r, row_data in enumerate(q2_data, 3):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
