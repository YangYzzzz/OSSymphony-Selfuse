"""
Reward Script: Merge cells A1:E1 on Summary sheet, center text, dark blue bg, white bold 16pt Calibri
Task ID: calc_ggf_011
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Cells A1:E1 are merged
  Component 2 (0.20): Text 'Annual Financial Summary' centered horizontally
  Component 3 (0.20): Dark blue background fill on A1
  Component 4 (0.30): White bold 16pt Calibri font on A1
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_011'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']
    cell = ws['A1']

    # Precondition: A1 must contain the expected text
    if cell.value != 'Annual Financial Summary':
        print(f"FAIL: A1 text mismatch — expected 'Annual Financial Summary', found {repr(cell.value)}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Cells A1:E1 are merged (0.30 points)
    # In initial_env there are no merged ranges; in golden_env A1:E1 is merged.
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        # Check that A1:E1 is among the merged ranges
        a1_e1_merged = False
        for mr in merged_ranges:
            if mr.upper() == 'A1:E1':
                a1_e1_merged = True
                break

        # Also verify B1-E1 are MergedCell objects (consequence of the merge)
        b1_merged = isinstance(ws['B1'], MergedCell)
        c1_merged = isinstance(ws['C1'], MergedCell)
        d1_merged = isinstance(ws['D1'], MergedCell)
        e1_merged = isinstance(ws['E1'], MergedCell)
        all_sub_merged = b1_merged and c1_merged and d1_merged and e1_merged

        if a1_e1_merged and all_sub_merged:
            print(f"PASS: Component 1 — A1:E1 merged correctly (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — merged ranges: {merged_ranges}, B1-E1 merged: {[b1_merged, c1_merged, d1_merged, e1_merged]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Text centered horizontally (0.20 points)
    # Initial has alignment.horizontal=None; golden has 'center'
    try:
        h_align = cell.alignment.horizontal
        if h_align == 'center':
            print(f"PASS: Component 2 — horizontal alignment is 'center' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — expected horizontal='center', found '{h_align}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Dark blue background fill (0.20 points)
    # Initial has fill_type=None (no fill); golden has solid fill with dark blue.
    # We accept a range of dark blue shades rather than one exact hex.
    try:
        fill_type = cell.fill.fill_type
        if fill_type != 'solid':
            print(f"FAIL: Component 3 — fill_type is '{fill_type}', expected 'solid'")
        else:
            fg_rgb = cell.fill.fgColor.rgb
            # The golden uses FF003366. Accept any dark blue:
            # Extract RGB components from ARGB string
            if fg_rgb and len(str(fg_rgb)) >= 6:
                rgb_str = str(fg_rgb)
                # Take last 6 chars as RGB
                rgb_hex = rgb_str[-6:]
                r = int(rgb_hex[0:2], 16)
                g = int(rgb_hex[2:4], 16)
                b = int(rgb_hex[4:6], 16)
                # Dark blue: blue channel dominant, RGB generally low except blue
                # Reasonable range: R<=80, G<=80, B>=50 (dark blues)
                is_dark_blue = (r <= 80 and g <= 80 and b >= 50) or rgb_hex.upper() == '003366'
                if is_dark_blue:
                    print(f"PASS: Component 3 — dark blue background fill ({rgb_hex}) (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 3 — fill color RGB={rgb_hex} (R={r},G={g},B={b}) is not dark blue")
            else:
                print(f"FAIL: Component 3 — could not parse fgColor: {fg_rgb}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: White bold 16pt Calibri font (0.30 points)
    # Initial: bold=False, size=11, theme color (black), name=Calibri.
    # Golden: bold=True, size=16, white, Calibri.
    # Only bold, size, and white color are task-introduced changes.
    # Font name 'Calibri' is already in initial, so it is a precondition gate, not a scoring item.
    try:
        comp4_score = 0.0

        # Precondition gate: font name must be Calibri (already present in initial)
        font_name = cell.font.name
        if not (font_name and font_name.lower() == 'calibri'):
            print(f"FAIL: Component 4 — font name='{font_name}', expected 'Calibri' (gate failed)")
        else:
            # 4a: Bold (0.10 points) — changes from False to True
            if cell.font.bold:
                comp4_score += 0.10
                print(f"  PASS: 4a — font bold=True")
            else:
                print(f"  FAIL: 4a — font bold={cell.font.bold}")

            # 4b: Size 16 (0.10 points) — changes from 11 to 16
            font_size = cell.font.size
            if font_size is not None and abs(float(font_size) - 16.0) < 0.5:
                comp4_score += 0.10
                print(f"  PASS: 4b — font size={font_size}")
            else:
                print(f"  FAIL: 4b — font size={font_size}, expected 16")

            # 4c: White font color (0.10 points) — changes from black/theme to white
            try:
                font_color = cell.font.color
                font_rgb = None
                if font_color is not None:
                    # Try to get RGB; may fail if it's a theme color
                    try:
                        raw_rgb = font_color.rgb
                        if isinstance(raw_rgb, str) and len(raw_rgb) >= 6:
                            font_rgb = raw_rgb
                    except (TypeError, ValueError):
                        pass
                if font_rgb:
                    rgb_hex = font_rgb[-6:]
                    r = int(rgb_hex[0:2], 16)
                    g = int(rgb_hex[2:4], 16)
                    b = int(rgb_hex[4:6], 16)
                    # White: all channels high (>=220)
                    if r >= 220 and g >= 220 and b >= 220:
                        comp4_score += 0.10
                        print(f"  PASS: 4c — font color is white ({rgb_hex})")
                    else:
                        print(f"  FAIL: 4c — font color RGB={rgb_hex} not white")
                else:
                    print(f"  FAIL: 4c — font color is theme/None (not explicit white)")
            except Exception as e:
                print(f"  ERROR: 4c — {e}")

        comp4_score = round(comp4_score, 4)
        if comp4_score > 0:
            print(f"PASS: Component 4 — font properties ({comp4_score} pts)")
            total_score += comp4_score
        else:
            print(f"FAIL: Component 4 — no font sub-checks passed")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
