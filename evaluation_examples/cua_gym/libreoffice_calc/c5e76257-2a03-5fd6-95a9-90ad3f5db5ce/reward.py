"""
Reward Script: AR Invoice Aging Report Formatting
Task ID: calc_fin_ar_invoice_aging_format_067
Domain: libreoffice_calc
Scoring:
  Component 1: Subtotal rows inserted for each customer group with SUM formulas (0.40 pts)
    - 0.20 for all 5 total rows present
    - 0.10 for all bold
    - 0.10 for all with SUM formula in E
  Component 2: E column formatted as currency (#,##0.00) (0.20 pts)
  Component 3: Row 1 headers bold and center-aligned (0.20 pts)
  Component 4: Page setup — landscape, letter, print area, row 1 repeat (0.20 pts)
Total: 1.0
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_ar_invoice_aging_format_067'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: AR_Report sheet must exist
    if 'AR_Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'AR_Report' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['AR_Report']

    # -------------------------------------------------------------------------
    # Component 1: Subtotal rows inserted for each customer group (0.40 pts)
    # Expect 5 total rows with customer name "XYZ Total" and SUM formula in col E
    # -------------------------------------------------------------------------
    try:
        expected_totals = {
            'Alpha Corp Total', 'Beta LLC Total', 'Gamma Inc Total',
            'Delta Co Total', 'Epsilon Ltd Total'
        }
        found_totals = {}
        found_bold_count = 0
        found_sum_count = 0

        for row in range(1, ws.max_row + 1):
            a_val = ws.cell(row=row, column=1).value
            if a_val and str(a_val).strip() in expected_totals:
                label = str(a_val).strip()
                e_cell = ws.cell(row=row, column=5)
                a_cell = ws.cell(row=row, column=1)
                found_totals[label] = row

                if a_cell.font.bold:
                    found_bold_count += 1

                e_val = e_cell.value
                if e_val and isinstance(e_val, str) and e_val.upper().startswith('=SUM('):
                    found_sum_count += 1

        # Sub-check 1a: all 5 total rows present (0.20 pts)
        if len(found_totals) == 5:
            print(f"PASS: Component 1a — All 5 customer subtotal rows found: {sorted(found_totals.keys())}")
            total_score += 0.20
        else:
            missing = sorted(expected_totals - set(found_totals.keys()))
            print(f"FAIL: Component 1a — Only {len(found_totals)}/5 total rows found. Missing: {missing}")

        # Sub-check 1b: all total rows are bold (0.10 pts) — only meaningful if rows exist
        if len(found_totals) > 0 and found_bold_count == len(found_totals):
            print(f"PASS: Component 1b — All {found_bold_count} found subtotal rows are bold")
            total_score += 0.10
        elif len(found_totals) > 0:
            print(f"FAIL: Component 1b — Only {found_bold_count}/{len(found_totals)} subtotal rows are bold")
        else:
            print("FAIL: Component 1b — No subtotal rows found to check for bold")

        # Sub-check 1c: all total rows have SUM formula (0.10 pts)
        if len(found_totals) > 0 and found_sum_count == len(found_totals):
            print(f"PASS: Component 1c — All {found_sum_count} found subtotal rows have SUM formula in col E")
            total_score += 0.10
        elif len(found_totals) > 0:
            print(f"FAIL: Component 1c — Only {found_sum_count}/{len(found_totals)} subtotal rows have SUM formula")
        else:
            print("FAIL: Component 1c — No subtotal rows found to check for SUM formula")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: E column formatted as currency '#,##0.00' (0.20 pts)
    # In initial file, the format is 'General'; task requires currency with commas
    # -------------------------------------------------------------------------
    try:
        max_row = ws.max_row
        formatted_count = 0
        total_e_rows = 0

        for row in range(2, max_row + 1):
            cell = ws.cell(row=row, column=5)
            if cell.value is not None:
                total_e_rows += 1
                fmt = cell.number_format or ''
                # Accept any format that includes comma and decimal notation
                if ',' in fmt and '0.00' in fmt:
                    formatted_count += 1

        if total_e_rows > 0 and formatted_count == total_e_rows:
            print(f"PASS: Component 2 — All {total_e_rows} E-column cells have currency format")
            total_score += 0.20
        elif total_e_rows > 0 and formatted_count > 0:
            print(f"PARTIAL: Component 2 — {formatted_count}/{total_e_rows} E-column cells have currency format")
            # Partial credit proportional to coverage
            total_score += round(0.20 * formatted_count / total_e_rows, 2)
        else:
            sample_fmt = repr(ws.cell(row=2, column=5).number_format) if ws.max_row >= 2 else 'N/A'
            print(f"FAIL: Component 2 — E column not formatted as currency. Sample format: {sample_fmt}")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Row 1 headers bold and center-aligned (0.20 pts)
    # In initial file, headers have bold=False and alignment=None
    # -------------------------------------------------------------------------
    try:
        bold_count = 0
        center_count = 0
        header_details = []

        for col in range(1, 6):
            cell = ws.cell(row=1, column=col)
            is_bold = cell.font.bold == True
            is_center = cell.alignment.horizontal == 'center'
            header_details.append(
                f"{get_column_letter(col)}1: bold={is_bold}, align={cell.alignment.horizontal}"
            )
            if is_bold:
                bold_count += 1
            if is_center:
                center_count += 1

        if bold_count == 5 and center_count == 5:
            print(f"PASS: Component 3 — All 5 header cells are bold and center-aligned")
            total_score += 0.20
        elif bold_count == 5:
            print(f"PARTIAL: Component 3 — Headers all bold but only {center_count}/5 center-aligned")
            total_score += 0.10
        elif center_count == 5:
            print(f"PARTIAL: Component 3 — Headers all center-aligned but only {bold_count}/5 bold")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — bold={bold_count}/5, centered={center_count}/5. Details: {header_details}")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Page setup — landscape, letter, print area, row 1 repeat (0.20 pts)
    # In initial file: orientation=None, paperSize=None, no print area, no repeat rows
    # -------------------------------------------------------------------------
    try:
        ps = ws.page_setup

        # Sub-check 4a: landscape orientation (0.05 pts)
        if ps.orientation == 'landscape':
            print(f"PASS: Component 4a — Landscape orientation set")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4a — Orientation: {repr(ps.orientation)} (expected 'landscape')")

        # Sub-check 4b: letter paper size (0.05 pts)
        if ps.paperSize == 1:
            print(f"PASS: Component 4b — Letter paper size set (paperSize=1)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4b — Paper size: {repr(ps.paperSize)} (expected 1 for Letter)")

        # Sub-check 4c: print area set (0.05 pts)
        print_area = ws.print_area
        if print_area and len(print_area) > 0:
            print(f"PASS: Component 4c — Print area configured: {print_area}")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4c — Print area not set (found: {repr(print_area)})")

        # Sub-check 4d: row 1 repeats on each page (0.05 pts)
        if ws.print_title_rows == '$1:$1':
            print(f"PASS: Component 4d — Row 1 repeat set (print_title_rows={ws.print_title_rows})")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4d — Row 1 not set to repeat (print_title_rows={repr(ws.print_title_rows)})")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
