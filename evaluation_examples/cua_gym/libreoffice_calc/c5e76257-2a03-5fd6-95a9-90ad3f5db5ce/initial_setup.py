"""
Initial Setup: AR Invoice Aging Report - Accounts Receivable spreadsheet
Task ID: calc_fin_ar_invoice_aging_format_067
Domain: libreoffice_calc

Creates an AR_Report sheet with 5 customer groups (79 invoice rows), no formatting,
no subtotals, no print area configuration.
"""

import os
from datetime import date
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_ar_invoice_aging_format_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AR_Report'

    # --- Row 1: Headers (plain, no formatting) ---
    headers = ['Customer', 'Invoice#', 'Invoice Date', 'Due Date', 'Amount']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Helper to write an invoice row
    def write_invoice(row, customer, inv_num, inv_date, due_date, amount):
        ws.cell(row=row, column=1, value=customer)
        ws.cell(row=row, column=2, value=inv_num)
        ws.cell(row=row, column=3, value=inv_date)
        ws.cell(row=row, column=4, value=due_date)
        ws.cell(row=row, column=5, value=amount)

    # --- Alpha Corp: rows 2-15 (14 invoices) ---
    alpha_data = [
        ('Alpha Corp', 'INV-10101', date(2025, 1, 5), date(2025, 2, 4), 3420.50),
        ('Alpha Corp', 'INV-10102', date(2025, 1, 12), date(2025, 2, 11), 1875.00),
        ('Alpha Corp', 'INV-10103', date(2025, 1, 18), date(2025, 2, 17), 5640.75),
        ('Alpha Corp', 'INV-10104', date(2025, 1, 22), date(2025, 2, 21), 2310.00),
        ('Alpha Corp', 'INV-10105', date(2025, 1, 28), date(2025, 2, 27), 4125.25),
        ('Alpha Corp', 'INV-10106', date(2025, 2, 3),  date(2025, 3, 5),  6890.00),
        ('Alpha Corp', 'INV-10107', date(2025, 2, 8),  date(2025, 3, 10), 1560.50),
        ('Alpha Corp', 'INV-10108', date(2025, 2, 14), date(2025, 3, 16), 3245.75),
        ('Alpha Corp', 'INV-10109', date(2025, 2, 19), date(2025, 3, 21), 7810.00),
        ('Alpha Corp', 'INV-10110', date(2025, 2, 25), date(2025, 3, 27), 2095.50),
        ('Alpha Corp', 'INV-10111', date(2025, 3, 3),  date(2025, 4, 2),  4430.25),
        ('Alpha Corp', 'INV-10112', date(2025, 3, 9),  date(2025, 4, 8),  1285.00),
        ('Alpha Corp', 'INV-10113', date(2025, 3, 15), date(2025, 4, 14), 5920.75),
        ('Alpha Corp', 'INV-10114', date(2025, 3, 20), date(2025, 4, 19), 3175.00),
    ]
    for i, row_data in enumerate(alpha_data, 2):
        write_invoice(i, *row_data)

    # --- Beta LLC: rows 16-28 (13 invoices) ---
    beta_data = [
        ('Beta LLC', 'INV-20201', date(2025, 1, 6),  date(2025, 2, 5),  9250.00),
        ('Beta LLC', 'INV-20202', date(2025, 1, 13), date(2025, 2, 12), 2840.50),
        ('Beta LLC', 'INV-20203', date(2025, 1, 20), date(2025, 2, 19), 6730.25),
        ('Beta LLC', 'INV-20204', date(2025, 1, 27), date(2025, 2, 26), 1450.00),
        ('Beta LLC', 'INV-20205', date(2025, 2, 4),  date(2025, 3, 6),  3895.75),
        ('Beta LLC', 'INV-20206', date(2025, 2, 10), date(2025, 3, 12), 8120.00),
        ('Beta LLC', 'INV-20207', date(2025, 2, 17), date(2025, 3, 19), 4360.50),
        ('Beta LLC', 'INV-20208', date(2025, 2, 24), date(2025, 3, 26), 2175.25),
        ('Beta LLC', 'INV-20209', date(2025, 3, 3),  date(2025, 4, 2),  7540.00),
        ('Beta LLC', 'INV-20210', date(2025, 3, 10), date(2025, 4, 9),  1680.50),
        ('Beta LLC', 'INV-20211', date(2025, 3, 17), date(2025, 4, 16), 5210.75),
        ('Beta LLC', 'INV-20212', date(2025, 3, 24), date(2025, 4, 23), 3405.00),
        ('Beta LLC', 'INV-20213', date(2025, 3, 28), date(2025, 4, 27), 1925.25),
    ]
    for i, row_data in enumerate(beta_data, 16):
        write_invoice(i, *row_data)

    # --- Gamma Inc: rows 29-45 (17 invoices) ---
    gamma_data = [
        ('Gamma Inc', 'INV-30301', date(2025, 1, 7),  date(2025, 2, 6),  4750.00),
        ('Gamma Inc', 'INV-30302', date(2025, 1, 10), date(2025, 2, 9),  8960.25),
        ('Gamma Inc', 'INV-30303', date(2025, 1, 15), date(2025, 2, 14), 2340.50),
        ('Gamma Inc', 'INV-30304', date(2025, 1, 21), date(2025, 2, 20), 6180.75),
        ('Gamma Inc', 'INV-30305', date(2025, 1, 28), date(2025, 2, 27), 1890.00),
        ('Gamma Inc', 'INV-30306', date(2025, 2, 5),  date(2025, 3, 7),  5425.50),
        ('Gamma Inc', 'INV-30307', date(2025, 2, 11), date(2025, 3, 13), 3670.25),
        ('Gamma Inc', 'INV-30308', date(2025, 2, 18), date(2025, 3, 20), 9015.00),
        ('Gamma Inc', 'INV-30309', date(2025, 2, 25), date(2025, 3, 27), 2560.75),
        ('Gamma Inc', 'INV-30310', date(2025, 3, 4),  date(2025, 4, 3),  7300.50),
        ('Gamma Inc', 'INV-30311', date(2025, 3, 11), date(2025, 4, 10), 4085.25),
        ('Gamma Inc', 'INV-30312', date(2025, 3, 18), date(2025, 4, 17), 1705.00),
        ('Gamma Inc', 'INV-30313', date(2025, 3, 21), date(2025, 4, 20), 6840.75),
        ('Gamma Inc', 'INV-30314', date(2025, 3, 25), date(2025, 4, 24), 3215.50),
        ('Gamma Inc', 'INV-30315', date(2025, 3, 28), date(2025, 4, 27), 8490.00),
        ('Gamma Inc', 'INV-30316', date(2025, 4, 2),  date(2025, 5, 2),  2095.25),
        ('Gamma Inc', 'INV-30317', date(2025, 4, 7),  date(2025, 5, 7),  5630.00),
    ]
    for i, row_data in enumerate(gamma_data, 29):
        write_invoice(i, *row_data)

    # --- Delta Co: rows 46-62 (17 invoices) ---
    delta_data = [
        ('Delta Co', 'INV-40401', date(2025, 1, 8),  date(2025, 2, 7),  6340.00),
        ('Delta Co', 'INV-40402', date(2025, 1, 14), date(2025, 2, 13), 2870.50),
        ('Delta Co', 'INV-40403', date(2025, 1, 20), date(2025, 2, 19), 4510.25),
        ('Delta Co', 'INV-40404', date(2025, 1, 26), date(2025, 2, 25), 9680.75),
        ('Delta Co', 'INV-40405', date(2025, 2, 2),  date(2025, 3, 4),  1395.00),
        ('Delta Co', 'INV-40406', date(2025, 2, 9),  date(2025, 3, 11), 7825.50),
        ('Delta Co', 'INV-40407', date(2025, 2, 16), date(2025, 3, 18), 3460.25),
        ('Delta Co', 'INV-40408', date(2025, 2, 23), date(2025, 3, 25), 5190.00),
        ('Delta Co', 'INV-40409', date(2025, 3, 2),  date(2025, 4, 1),  8050.75),
        ('Delta Co', 'INV-40410', date(2025, 3, 9),  date(2025, 4, 8),  2635.50),
        ('Delta Co', 'INV-40411', date(2025, 3, 16), date(2025, 4, 15), 6145.25),
        ('Delta Co', 'INV-40412', date(2025, 3, 23), date(2025, 4, 22), 1870.00),
        ('Delta Co', 'INV-40413', date(2025, 3, 30), date(2025, 4, 29), 4380.75),
        ('Delta Co', 'INV-40414', date(2025, 4, 4),  date(2025, 5, 4),  9920.50),
        ('Delta Co', 'INV-40415', date(2025, 4, 9),  date(2025, 5, 9),  3250.25),
        ('Delta Co', 'INV-40416', date(2025, 4, 14), date(2025, 5, 14), 7610.00),
        ('Delta Co', 'INV-40417', date(2025, 4, 18), date(2025, 5, 18), 2195.50),
    ]
    for i, row_data in enumerate(delta_data, 46):
        write_invoice(i, *row_data)

    # --- Epsilon Ltd: rows 63-80 (18 invoices) ---
    epsilon_data = [
        ('Epsilon Ltd', 'INV-50501', date(2025, 1, 9),  date(2025, 2, 8),  5890.00),
        ('Epsilon Ltd', 'INV-50502', date(2025, 1, 15), date(2025, 2, 14), 3120.25),
        ('Epsilon Ltd', 'INV-50503', date(2025, 1, 21), date(2025, 2, 20), 7460.75),
        ('Epsilon Ltd', 'INV-50504', date(2025, 1, 28), date(2025, 2, 27), 1940.50),
        ('Epsilon Ltd', 'INV-50505', date(2025, 2, 4),  date(2025, 3, 6),  4280.00),
        ('Epsilon Ltd', 'INV-50506', date(2025, 2, 11), date(2025, 3, 13), 8750.25),
        ('Epsilon Ltd', 'INV-50507', date(2025, 2, 18), date(2025, 3, 20), 2365.75),
        ('Epsilon Ltd', 'INV-50508', date(2025, 2, 25), date(2025, 3, 27), 6015.50),
        ('Epsilon Ltd', 'INV-50509', date(2025, 3, 4),  date(2025, 4, 3),  3840.00),
        ('Epsilon Ltd', 'INV-50510', date(2025, 3, 11), date(2025, 4, 10), 9175.25),
        ('Epsilon Ltd', 'INV-50511', date(2025, 3, 18), date(2025, 4, 17), 1650.75),
        ('Epsilon Ltd', 'INV-50512', date(2025, 3, 25), date(2025, 4, 24), 5430.50),
        ('Epsilon Ltd', 'INV-50513', date(2025, 4, 1),  date(2025, 5, 1),  7895.00),
        ('Epsilon Ltd', 'INV-50514', date(2025, 4, 7),  date(2025, 5, 7),  2740.25),
        ('Epsilon Ltd', 'INV-50515', date(2025, 4, 13), date(2025, 5, 13), 4560.75),
        ('Epsilon Ltd', 'INV-50516', date(2025, 4, 18), date(2025, 5, 18), 8310.00),
        ('Epsilon Ltd', 'INV-50517', date(2025, 4, 23), date(2025, 5, 23), 1285.50),
        ('Epsilon Ltd', 'INV-50518', date(2025, 4, 28), date(2025, 5, 28), 6920.75),
    ]
    for i, row_data in enumerate(epsilon_data, 63):
        write_invoice(i, *row_data)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  AR_Report rows: 1 header + 79 data rows (rows 2-80)')


create_initial()
