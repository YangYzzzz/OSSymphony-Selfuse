"""
Initial Setup: Project Tracker - Email Data Extraction Task
Task ID: osworld_multi_apps_email_data_007
Domain: libreoffice_calc (multi-app: Thunderbird + LibreOffice Calc)

Creates:
  1. /home/user/project_tracker.ods  -- spreadsheet with headers only (no data rows)
  2. Thunderbird 'Project Alpha' folder with 5 emails containing 'Progress: X%'
  3. Opens Thunderbird and LibreOffice Calc with the tracker file
"""

import os
import shlex
import subprocess
import time
import glob

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_data_007'
OUTPUT_ODS = f'{WORKDIR}/project_tracker.ods'

# ---------------------------------------------------------------------------
# Email data (5 emails, unsorted, each contains "Progress: X%")
# ---------------------------------------------------------------------------
EMAILS = [
    {
        "from_name": "Alice Morgan",
        "from_addr": "alice.morgan@techcorp.com",
        "date_str": "Mon, 10 Mar 2025 09:15:00 +0000",
        "subject": "Project Alpha - Weekly Update",
        "body": (
            "Hi Team,\n\n"
            "Here is the weekly status update for Project Alpha.\n"
            "Progress: 35%\n\n"
            "We completed the initial design phase and are now moving into development.\n"
            "The team is on track to meet the Q2 milestone.\n\n"
            "Best regards,\nAlice Morgan\nProject Lead"
        ),
    },
    {
        "from_name": "Bob Stevenson",
        "from_addr": "bob.stevenson@techcorp.com",
        "date_str": "Mon, 17 Mar 2025 10:30:00 +0000",
        "subject": "Project Alpha - Status Report Week 11",
        "body": (
            "Hello,\n\n"
            "Sharing this week's progress for Project Alpha.\n"
            "Progress: 58%\n\n"
            "Backend API integration is now complete. Frontend components are being finalized.\n"
            "No blockers at this time.\n\n"
            "Regards,\nBob Stevenson\nDev Lead"
        ),
    },
    {
        "from_name": "Carol Lee",
        "from_addr": "carol.lee@techcorp.com",
        "date_str": "Mon, 03 Mar 2025 08:45:00 +0000",
        "subject": "Project Alpha - Kickoff Summary",
        "body": (
            "Team,\n\n"
            "Following our kickoff meeting, here is where we stand.\n"
            "Progress: 22%\n\n"
            "Requirements gathering is complete. Architecture review scheduled for next week.\n"
            "Please review attached documents before the next meeting.\n\n"
            "Thanks,\nCarol Lee\nBusiness Analyst"
        ),
    },
    {
        "from_name": "David Park",
        "from_addr": "david.park@techcorp.com",
        "date_str": "Mon, 24 Mar 2025 11:00:00 +0000",
        "subject": "Project Alpha - Sprint 4 Completion",
        "body": (
            "Hi all,\n\n"
            "Sprint 4 has concluded successfully for Project Alpha.\n"
            "Progress: 75%\n\n"
            "Testing phase begins next Monday. UAT sessions are scheduled for the following week.\n"
            "Please ensure all feature branches are merged by EOD Friday.\n\n"
            "Best,\nDavid Park\nScrum Master"
        ),
    },
    {
        "from_name": "Alice Morgan",
        "from_addr": "alice.morgan@techcorp.com",
        "date_str": "Mon, 31 Mar 2025 14:20:00 +0000",
        "subject": "Project Alpha - Final Stretch Update",
        "body": (
            "Team,\n\n"
            "Exciting news on Project Alpha - we are almost done!\n"
            "Progress: 91%\n\n"
            "All critical features are implemented and tested. Final sign-off expected next week.\n"
            "Great work everyone - we are ahead of schedule.\n\n"
            "Best regards,\nAlice Morgan\nProject Lead"
        ),
    },
]


# ---------------------------------------------------------------------------
# Helper: launch GUI app on VM display
# ---------------------------------------------------------------------------
def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch a GUI app on DISPLAY=:0, non-blocking."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# ---------------------------------------------------------------------------
# Step 1: Create project_tracker.ods with headers only (no data rows)
# ---------------------------------------------------------------------------
def create_tracker_ods():
    """Create ODS spreadsheet with Sender, Date, Progress headers only."""
    # Write an ODS file directly using odfpy
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties, TableCellProperties
        import odf.style

        doc = OpenDocumentSpreadsheet()

        # Create header style - plain, no bold needed (task asks agent to fill data)
        header_style = Style(name="HeaderStyle", family="table-cell")
        doc.styles.addElement(header_style)

        table = Table(name="Sheet1")

        # Header row: Sender | Date | Progress
        header_row = TableRow()
        for header_text in ["Sender", "Date", "Progress"]:
            cell = TableCell(valuetype="string")
            cell.addElement(P(text=header_text))
            header_row.addElement(cell)
        # Column D header empty (D2 will hold the formula)
        empty_cell = TableCell(valuetype="string")
        empty_cell.addElement(P(text=""))
        header_row.addElement(empty_cell)

        table.addElement(header_row)
        doc.spreadsheet.addElement(table)
        doc.save(OUTPUT_ODS)
        print(f"Created ODS file: {OUTPUT_ODS}")
        return True
    except ImportError:
        print("odfpy not available, using xlsx + conversion fallback")
        return False


def create_tracker_xlsx_fallback():
    """Fallback: create xlsx then convert to ods via LibreOffice."""
    import openpyxl
    xlsx_path = f'{WORKDIR}/project_tracker.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # Headers only
    ws.cell(row=1, column=1, value="Sender")
    ws.cell(row=1, column=2, value="Date")
    ws.cell(row=1, column=3, value="Progress")
    wb.save(xlsx_path)
    print(f"Created XLSX: {xlsx_path}")
    return xlsx_path


# ---------------------------------------------------------------------------
# Step 2: Set up Thunderbird 'Project Alpha' folder with 5 emails
# ---------------------------------------------------------------------------
def setup_thunderbird_emails():
    """
    Create 'Project Alpha' mbox folder inside Thunderbird's local mail.
    Thunderbird stores local folders as mbox files in the profile directory.
    """
    # Find Thunderbird profile directory
    tb_base = os.path.expanduser('~/.thunderbird')
    if not os.path.exists(tb_base):
        print("WARNING: Thunderbird directory not found, creating minimal structure")
        os.makedirs(tb_base, exist_ok=True)

    # Look for existing profile
    profile_dir = None
    ini_path = os.path.join(tb_base, 'profiles.ini')
    if os.path.exists(ini_path):
        with open(ini_path, 'r') as f:
            for line in f:
                if line.startswith('Path='):
                    rel_path = line.strip().split('=', 1)[1]
                    candidate = os.path.join(tb_base, rel_path)
                    if os.path.exists(candidate):
                        profile_dir = candidate
                        break

    if profile_dir is None:
        # Find any .default or .default-release profile directory
        patterns = [
            os.path.join(tb_base, '*.default'),
            os.path.join(tb_base, '*.default-release'),
            os.path.join(tb_base, '*.default-esr'),
        ]
        for pat in patterns:
            matches = glob.glob(pat)
            if matches:
                profile_dir = matches[0]
                break

    if profile_dir is None:
        # Create a default profile
        import uuid
        profile_name = f"{uuid.uuid4().hex[:8]}.default"
        profile_dir = os.path.join(tb_base, profile_name)
        os.makedirs(profile_dir, exist_ok=True)
        # Write profiles.ini
        with open(ini_path, 'w') as f:
            f.write("[General]\nStartWithLastProfile=1\n\n[Profile0]\nName=default\nIsRelative=1\nPath={}\nDefault=1\n".format(profile_name))
        print(f"Created new Thunderbird profile: {profile_dir}")
    else:
        print(f"Found Thunderbird profile: {profile_dir}")

    # Local mail directory (local folders are in Mail/Local Folders/)
    local_mail_dir = os.path.join(profile_dir, 'Mail', 'Local Folders')
    os.makedirs(local_mail_dir, exist_ok=True)

    # Create 'Project Alpha' mbox file
    mbox_path = os.path.join(local_mail_dir, 'Project Alpha')
    mbox_sbd = os.path.join(local_mail_dir, 'Project Alpha.sbd')

    # Build mbox content
    mbox_content = _build_mbox()

    with open(mbox_path, 'w', encoding='utf-8') as f:
        f.write(mbox_content)
    print(f"Created mbox: {mbox_path}")

    # Create .msf index file (Thunderbird will regenerate it on open)
    msf_path = mbox_path + '.msf'
    with open(msf_path, 'w') as f:
        f.write('// <!-- <mdb:mork:z v="1.4"/> -->\n')
    print(f"Created msf: {msf_path}")

    # Ensure the folder is listed in Thunderbird's panacea.dat or prefs.js
    _update_thunderbird_prefs(profile_dir, local_mail_dir)

    return mbox_path


def _build_mbox():
    """Build mbox format string for all 5 emails."""
    lines = []
    for email in EMAILS:
        # mbox separator line
        lines.append(f'From {email["from_addr"]} {email["date_str"]}')
        lines.append(f'From: {email["from_name"]} <{email["from_addr"]}>')
        lines.append(f'To: project-alpha-team@techcorp.com')
        lines.append(f'Date: {email["date_str"]}')
        lines.append(f'Subject: {email["subject"]}')
        lines.append(f'Message-ID: <{email["from_addr"].split("@")[0]}.{email["date_str"].replace(",","").replace(" ","").replace(":","")[3:15]}@techcorp.com>')
        lines.append('MIME-Version: 1.0')
        lines.append('Content-Type: text/plain; charset=UTF-8')
        lines.append('')
        # Body
        for body_line in email['body'].split('\n'):
            # mbox "From " escaping
            if body_line.startswith('From '):
                body_line = '>' + body_line
            lines.append(body_line)
        lines.append('')  # blank line between messages
    return '\n'.join(lines)


def _update_thunderbird_prefs(profile_dir, local_mail_dir):
    """Update Thunderbird prefs.js to reference local mail directory."""
    prefs_path = os.path.join(profile_dir, 'prefs.js')
    local_folder_pref = f'user_pref("mail.server.server1.directory", "{local_mail_dir}");'
    server_type_pref = 'user_pref("mail.server.server1.type", "none");'
    server_name_pref = 'user_pref("mail.server.server1.name", "Local Folders");'
    server_hostname_pref = 'user_pref("mail.server.server1.hostname", "Local Folders");'
    account_pref = 'user_pref("mail.account.account1.server", "server1");'
    accounts_pref = 'user_pref("mail.accountmanager.accounts", "account1");'
    default_account_pref = 'user_pref("mail.accountmanager.defaultaccount", "account1");'

    new_prefs = [
        local_folder_pref,
        server_type_pref,
        server_name_pref,
        server_hostname_pref,
        account_pref,
        accounts_pref,
        default_account_pref,
    ]

    existing_content = ''
    if os.path.exists(prefs_path):
        with open(prefs_path, 'r') as f:
            existing_content = f.read()

    additions = []
    for pref in new_prefs:
        key = pref.split('"')[1]
        if key not in existing_content:
            additions.append(pref)

    if additions:
        with open(prefs_path, 'a') as f:
            f.write('\n// Added by initial_setup.py\n')
            for p in additions:
                f.write(p + '\n')
        print(f"Updated prefs.js with {len(additions)} entries")
    else:
        print("prefs.js already configured")


# ---------------------------------------------------------------------------
# Step 3: Convert xlsx to ods if needed (run on VM via LibreOffice headless)
# ---------------------------------------------------------------------------
def convert_to_ods_if_needed():
    """If ODS creation failed, convert xlsx to ods using LibreOffice headless."""
    if not os.path.exists(OUTPUT_ODS):
        xlsx_path = f'{WORKDIR}/project_tracker.xlsx'
        if os.path.exists(xlsx_path):
            env = os.environ.copy()
            env["DISPLAY"] = ":0"
            cmd = f'libreoffice --headless --convert-to ods --outdir {WORKDIR} {xlsx_path}'
            result = subprocess.run(
                shlex.split(cmd),
                env=env,
                capture_output=True,
                text=True,
                timeout=60,
            )
            if result.returncode == 0 and os.path.exists(OUTPUT_ODS):
                os.remove(xlsx_path)
                print(f"Converted xlsx to ods: {OUTPUT_ODS}")
            else:
                print(f"Conversion failed: {result.stderr}")
                # Keep the xlsx and rename it
                os.rename(xlsx_path, OUTPUT_ODS.replace('.ods', '.xlsx'))
                print("Kept as xlsx instead")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    print("=== Initial Setup: osworld_multi_apps_email_data_007 ===")

    # 1. Create ODS tracker file
    ods_created = create_tracker_ods()
    if not ods_created:
        create_tracker_xlsx_fallback()
        convert_to_ods_if_needed()

    # Verify ODS exists
    if os.path.exists(OUTPUT_ODS):
        print(f"[OK] ODS file exists: {OUTPUT_ODS} ({os.path.getsize(OUTPUT_ODS)} bytes)")
    else:
        print(f"[WARN] ODS file not found at {OUTPUT_ODS}")

    # 2. Set up Thunderbird emails
    mbox_path = setup_thunderbird_emails()
    print(f"[OK] Thunderbird mbox: {mbox_path}")

    # 3. Launch Thunderbird (primary app)
    print("Launching Thunderbird...")
    launch_gui('thunderbird', delay_sec=3.0)

    # 4. Launch LibreOffice Calc with the tracker file
    print("Launching LibreOffice Calc with project_tracker.ods...")
    launch_gui(f'libreoffice --calc "{OUTPUT_ODS}"', delay_sec=2.0)

    print("GUI_READY: launched Thunderbird and LibreOffice Calc with DISPLAY=:0")
    print("=== Initial setup complete ===")


main()
