"""
Initial Setup: Class Participation Tracker - raw data only
Task ID: calc_gpm_082
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

STUDENTS = [
    'Sarah Chen',
    'Marcus Johnson',
    'Priya Patel',
    'James O\'Brien',
    'Fatima Al-Rashid',
    'David Kim',
    'Elena Rodriguez',
    'Tyler Washington',
    'Aisha Mbeki',
    'Ryan Nakamura',
    'Sophie Laurent',
    'Carlos Mendez',
    'Hannah Fischer',
    'Omar Hassan',
    'Lily Chang',
]

# Weekly participation scores (0-5) for 15 students x 10 weeks
SCORES = [
    [5, 4, 5, 3, 5, 4, 5, 5, 4, 5],   # Sarah Chen
    [3, 3, 4, 4, 3, 2, 4, 3, 4, 3],   # Marcus Johnson
    [4, 5, 4, 5, 4, 5, 5, 4, 5, 4],   # Priya Patel
    [2, 3, 2, 3, 3, 2, 3, 2, 3, 3],   # James O'Brien
    [5, 5, 4, 5, 5, 5, 4, 5, 5, 5],   # Fatima Al-Rashid
    [3, 2, 3, 3, 4, 3, 3, 4, 3, 3],   # David Kim
    [4, 4, 5, 4, 4, 3, 4, 4, 5, 4],   # Elena Rodriguez
    [1, 2, 1, 0, 2, 1, 2, 1, 0, 1],   # Tyler Washington
    [4, 3, 4, 4, 5, 4, 4, 3, 4, 5],   # Aisha Mbeki
    [3, 4, 3, 4, 3, 4, 3, 4, 3, 4],   # Ryan Nakamura
    [5, 5, 5, 4, 5, 5, 5, 5, 4, 5],   # Sophie Laurent
    [2, 1, 2, 3, 2, 1, 2, 2, 3, 2],   # Carlos Mendez
    [4, 4, 3, 4, 4, 5, 4, 4, 3, 4],   # Hannah Fischer
    [0, 1, 0, 1, 0, 1, 0, 1, 0, 1],   # Omar Hassan
    [5, 4, 5, 5, 4, 5, 5, 4, 5, 5],   # Lily Chang
]


def launch_gui(command: str, delay_sec: float = 1.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Participation'

    # Row 1: plain title (no merge, no formatting - task is to build the tracker)
    ws.cell(row=1, column=1, value='Class Participation Tracker - Fall 2025')

    # Row 2: plain headers
    headers = ['Student'] + [f'W{i}' for i in range(1, 11)] + ['Average']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # Rows 3-17: student names and scores (raw, unformatted)
    for r, (name, scores) in enumerate(zip(STUDENTS, SCORES), 3):
        ws.cell(row=r, column=1, value=name)
        for c, score in enumerate(scores, 2):
            ws.cell(row=r, column=c, value=score)
        # L column intentionally empty - task is to add AVERAGE formulas

    # Set basic column widths for readability
    ws.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
        ws.column_dimensions[col_letter].width = 6
    ws.column_dimensions['L'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
