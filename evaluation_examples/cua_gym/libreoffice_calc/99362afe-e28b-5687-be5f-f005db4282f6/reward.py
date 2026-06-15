"""
Reward Script: Class Participation Tracker with weekly scores and semester trend visualization
Task ID: calc_gpm_082
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15) - Merged title A1:L1 with formatting (bold, 14pt, centered, maroon fill)
  Component 2 (0.10) - Row 2 headers bold with maroon fill
  Component 3 (0.15) - AVERAGE formulas in L3:L17 with 1 decimal format
  Component 4 (0.15) - Row 19 Class Average with AVERAGE formulas and gray fill/bold/top border
  Component 5 (0.10) - Row 20 Participation Rate with COUNTIF formulas and percentage format
  Component 6 (0.15) - Conditional formatting rules (score grid + L column + data bars)
  Component 7 (0.20) - Line chart present with correct title
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_082'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Participation' not in wb.sheetnames:
        print("FAIL: 'Participation' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Participation']

    # Component 1: Merged title A1:L1 with formatting (0.15 points)
    # Initial has NO merge, NO bold on A1, NO fill. Golden has merge + bold + 14pt + centered + maroon fill.
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'L1' in r for r in merged_ranges)
        c1 = ws['A1']
        has_bold = c1.font.bold is True
        has_size_14 = c1.font.size is not None and abs(c1.font.size - 14) < 1
        has_center = c1.alignment.horizontal == 'center'
        # Maroon fill: FF800000
        try:
            fill_rgb = c1.fill.fgColor.rgb
            has_maroon_fill = fill_rgb is not None and '800000' in str(fill_rgb)
        except Exception:
            has_maroon_fill = False

        checks_passed = sum([has_merge, has_bold, has_size_14, has_center, has_maroon_fill])
        if checks_passed >= 4:
            print(f"PASS: Component 1 - Title merged & formatted ({checks_passed}/5 sub-checks) (0.15 pts)")
            total_score += 0.15
        elif checks_passed >= 2:
            partial = 0.08
            print(f"PARTIAL: Component 1 - Title partially formatted ({checks_passed}/5 sub-checks) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Title not properly formatted (merge={has_merge}, bold={has_bold}, size14={has_size_14}, center={has_center}, maroon={has_maroon_fill})")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Row 2 headers bold with maroon fill and white text (0.10 points)
    # Initial has NO bold, NO fill on row 2. Golden has bold + maroon fill + white text.
    try:
        header_ok_count = 0
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
            cell = ws[f'{col_letter}2']
            is_bold = cell.font.bold is True
            try:
                fill_rgb = cell.fill.fgColor.rgb
                has_fill = fill_rgb is not None and '800000' in str(fill_rgb)
            except Exception:
                has_fill = False
            if is_bold and has_fill:
                header_ok_count += 1

        if header_ok_count >= 10:
            print(f"PASS: Component 2 - Row 2 headers bold with maroon fill ({header_ok_count}/12) (0.10 pts)")
            total_score += 0.10
        elif header_ok_count >= 5:
            print(f"PARTIAL: Component 2 - Row 2 headers partially formatted ({header_ok_count}/12) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 - Row 2 headers not formatted ({header_ok_count}/12)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: AVERAGE formulas in L3:L17 with 0.0 number format (0.15 points)
    # Initial has L3:L17 = None. Golden has =AVERAGE(B:K) formulas.
    try:
        formula_count = 0
        format_count = 0
        for row in range(3, 18):
            val = ws.cell(row=row, column=12).value
            nf = ws.cell(row=row, column=12).number_format
            if isinstance(val, str) and 'AVERAGE' in val.upper():
                formula_count += 1
            if nf is not None and '0.0' in str(nf) and '%' not in str(nf):
                format_count += 1

        if formula_count >= 13:
            pts = 0.10
            print(f"PASS: Component 3a - AVERAGE formulas in L3:L17 ({formula_count}/15) (0.10 pts)")
            total_score += pts
            if format_count >= 10:
                print(f"PASS: Component 3b - L column 1-decimal format ({format_count}/15) (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 3b - L column format ({format_count}/15 have 0.0 format)")
        else:
            print(f"FAIL: Component 3 - AVERAGE formulas in L column ({formula_count}/15)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Row 19 - Class Average with AVERAGE formulas, gray fill, bold, top border (0.15 points)
    # Initial has row 19 empty. Golden has "Class Average" label + AVERAGE formulas + formatting.
    try:
        a19_val = ws['A19'].value
        has_label = a19_val is not None and 'average' in str(a19_val).lower()

        formula_count_19 = 0
        for col in range(2, 12):  # B to K
            val = ws.cell(row=19, column=col).value
            if isinstance(val, str) and 'AVERAGE' in val.upper():
                formula_count_19 += 1

        has_formulas = formula_count_19 >= 8

        # Check styling: bold + gray fill + top border
        a19_bold = ws['A19'].font.bold is True
        try:
            fill_19 = ws['B19'].fill.fgColor.rgb
            has_gray_fill = fill_19 is not None and 'D9D9D9' in str(fill_19)
        except Exception:
            has_gray_fill = False
        has_top_border = ws['A19'].border.top.style is not None

        if has_label and has_formulas:
            pts = 0.10
            print(f"PASS: Component 4a - Row 19 Class Average with formulas ({formula_count_19}/10) (0.10 pts)")
            total_score += pts
            if a19_bold and (has_gray_fill or has_top_border):
                print(f"PASS: Component 4b - Row 19 formatted (bold={a19_bold}, gray={has_gray_fill}, border={has_top_border}) (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 4b - Row 19 formatting (bold={a19_bold}, gray={has_gray_fill}, border={has_top_border})")
        else:
            print(f"FAIL: Component 4 - Row 19 (label={has_label}, formulas={formula_count_19}/10)")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Row 20 - Participation Rate with COUNTIF formulas and percentage format (0.10 points)
    # Initial has row 20 empty. Golden has COUNTIF formulas and % format.
    try:
        a20_val = ws['A20'].value
        has_label_20 = a20_val is not None and 'participation' in str(a20_val).lower()

        countif_count = 0
        pct_count = 0
        for col in range(2, 12):  # B to K
            val = ws.cell(row=20, column=col).value
            nf = ws.cell(row=20, column=col).number_format
            if isinstance(val, str) and 'COUNTIF' in val.upper():
                countif_count += 1
            if nf is not None and '%' in str(nf):
                pct_count += 1

        if has_label_20 and countif_count >= 8:
            pts = 0.07
            print(f"PASS: Component 5a - Row 20 Participation Rate with COUNTIF ({countif_count}/10) (0.07 pts)")
            total_score += pts
            if pct_count >= 8:
                print(f"PASS: Component 5b - Row 20 percentage format ({pct_count}/10) (0.03 pts)")
                total_score += 0.03
            else:
                print(f"FAIL: Component 5b - Row 20 percentage format ({pct_count}/10)")
        else:
            print(f"FAIL: Component 5 - Row 20 (label={has_label_20}, countif={countif_count}/10)")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Conditional formatting rules (0.15 points)
    # Initial has 0 CF rules. Golden has 3 ranges with rules.
    try:
        cf_list = list(ws.conditional_formatting)
        cf_count = len(cf_list)

        total_rules = sum(len(cf.rules) for cf in cf_list)

        if cf_count >= 3 and total_rules >= 10:
            print(f"PASS: Component 6 - Conditional formatting ({cf_count} ranges, {total_rules} rules) (0.15 pts)")
            total_score += 0.15
        elif cf_count >= 2 and total_rules >= 6:
            print(f"PARTIAL: Component 6 - Conditional formatting ({cf_count} ranges, {total_rules} rules) (0.10 pts)")
            total_score += 0.10
        elif cf_count >= 1:
            print(f"PARTIAL: Component 6 - Some conditional formatting ({cf_count} ranges, {total_rules} rules) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 - No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Line chart with title "Weekly Participation Trend" (0.20 points)
    # Initial has 0 charts. Golden has 1 line chart.
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            is_line = type(chart).__name__ == 'LineChart'

            # Extract chart title text from openpyxl Title object
            chart_title_text = None
            try:
                t_obj = chart.title
                if t_obj is not None:
                    # Try string first
                    if isinstance(t_obj, str):
                        chart_title_text = t_obj
                    # Title object with rich text
                    elif hasattr(t_obj, 'tx') and t_obj.tx is not None:
                        if hasattr(t_obj.tx, 'rich') and t_obj.tx.rich is not None:
                            parts = []
                            for p in t_obj.tx.rich.p:
                                if hasattr(p, 'r') and p.r:
                                    for r in p.r:
                                        if hasattr(r, 't') and r.t:
                                            parts.append(r.t)
                            if parts:
                                chart_title_text = ''.join(parts)
                    # Fallback: try .text attribute
                    if chart_title_text is None and hasattr(t_obj, 'text'):
                        txt = t_obj.text
                        if isinstance(txt, str):
                            chart_title_text = txt
            except Exception:
                pass

            has_correct_title = chart_title_text is not None and 'participation' in chart_title_text.lower() and 'trend' in chart_title_text.lower()

            if is_line and has_correct_title:
                print(f"PASS: Component 7 - Line chart with title '{chart_title_text}' (0.20 pts)")
                total_score += 0.20
            elif is_line:
                print(f"PARTIAL: Component 7 - Line chart found but title mismatch (title='{chart_title_text}') (0.12 pts)")
                total_score += 0.12
            elif has_correct_title:
                print(f"PARTIAL: Component 7 - Chart found with correct title but not line type ({type(chart).__name__}) (0.12 pts)")
                total_score += 0.12
            else:
                print(f"PARTIAL: Component 7 - Chart found but wrong type and title (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 7 - No charts found")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
