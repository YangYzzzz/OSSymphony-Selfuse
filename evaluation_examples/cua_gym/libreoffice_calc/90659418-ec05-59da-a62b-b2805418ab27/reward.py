"""
Reward Script: Create 'Archive_2023' sheet, move data from 'Old Data', delete 'Old Data'
Task ID: calc_gsi_094
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): 'Archive_2023' sheet exists in the workbook
  Component 2 (0.3): 'Old Data' sheet has been deleted
  Component 3 (0.4): Data from 'Old Data' is preserved in 'Archive_2023' (spot-check)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_094'


def persist_app_state(domain: str):
    """Best-effort save of any unsaved GUI edits."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames
    print(f"INFO: Sheet names found: {sheet_names}")

    # Component 1: 'Archive_2023' sheet exists (0.3 points)
    try:
        if 'Archive_2023' in sheet_names:
            print(f"PASS: Component 1 — 'Archive_2023' sheet exists (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — 'Archive_2023' sheet not found in {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Old Data' sheet has been deleted (0.3 points)
    try:
        if 'Old Data' not in sheet_names:
            print(f"PASS: Component 2 — 'Old Data' sheet has been deleted (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — 'Old Data' sheet still exists")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Data preserved in 'Archive_2023' (0.4 points)
    # Verify key data points that were originally in 'Old Data':
    #   A1='Employee', B1='Department', ..., F1='Performance Rating'
    #   A2='Sarah Chen', D2=92000, F2='Exceeds'
    #   A13='Carlos Mendez', D13=74000, F13='Needs Improvement'
    #   Total: 13 rows x 6 cols, headers + 12 data rows
    try:
        if 'Archive_2023' not in sheet_names:
            print(f"FAIL: Component 3 — Cannot check data; 'Archive_2023' does not exist")
        else:
            ws = wb['Archive_2023']
            checks_passed = 0
            checks_total = 8

            # Check headers
            if ws['A1'].value == 'Employee':
                checks_passed += 1
            else:
                print(f"  DETAIL: A1 expected 'Employee', got '{ws['A1'].value}'")

            if ws['F1'].value == 'Performance Rating':
                checks_passed += 1
            else:
                print(f"  DETAIL: F1 expected 'Performance Rating', got '{ws['F1'].value}'")

            # Check first data row
            if ws['A2'].value == 'Sarah Chen':
                checks_passed += 1
            else:
                print(f"  DETAIL: A2 expected 'Sarah Chen', got '{ws['A2'].value}'")

            if ws['D2'].value == 92000:
                checks_passed += 1
            else:
                print(f"  DETAIL: D2 expected 92000, got '{ws['D2'].value}'")

            # Check middle data row
            if ws['A7'].value == 'David Kim':
                checks_passed += 1
            else:
                print(f"  DETAIL: A7 expected 'David Kim', got '{ws['A7'].value}'")

            if ws['E7'].value == 9800:
                checks_passed += 1
            else:
                print(f"  DETAIL: E7 expected 9800, got '{ws['E7'].value}'")

            # Check last data row
            if ws['A13'].value == 'Carlos Mendez':
                checks_passed += 1
            else:
                print(f"  DETAIL: A13 expected 'Carlos Mendez', got '{ws['A13'].value}'")

            if ws['F13'].value == 'Needs Improvement':
                checks_passed += 1
            else:
                print(f"  DETAIL: F13 expected 'Needs Improvement', got '{ws['F13'].value}'")

            # Award proportional credit
            component3_score = 0.4 * (checks_passed / checks_total)
            if checks_passed == checks_total:
                print(f"PASS: Component 3 — All {checks_total} data checks passed ({component3_score:.2f} pts)")
                total_score += component3_score
            elif checks_passed > 0:
                print(f"PARTIAL: Component 3 — {checks_passed}/{checks_total} data checks passed ({component3_score:.2f} pts)")
                total_score += component3_score
            else:
                print(f"FAIL: Component 3 — 0/{checks_total} data checks passed")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
