"""
Initial Setup: Create workbook with 'Old Data' sheet to be archived
Task ID: calc_gsi_094
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Sales (current data) ---
    ws_sales = wb.active
    ws_sales.title = 'Sales'

    sales_headers = ['Product', 'Region', 'Q1 Revenue', 'Q2 Revenue', 'Q3 Revenue', 'Q4 Revenue', 'Total']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(sales_headers, 1):
        cell = ws_sales.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    sales_data = [
        ['Laptop Pro 15"', 'North America', 125400, 138200, 142800, 156300, None],
        ['Wireless Mouse X1', 'Europe', 34500, 36800, 38200, 41500, None],
        ['USB-C Hub Ultra', 'Asia Pacific', 28700, 31200, 33600, 35800, None],
        ['Mechanical Keyboard K7', 'North America', 45200, 48900, 51300, 54700, None],
        ['Monitor 27" 4K', 'Europe', 89300, 92100, 96400, 101200, None],
        ['Webcam HD Pro', 'Asia Pacific', 15600, 17200, 18900, 20300, None],
        ['Docking Station D3', 'North America', 62800, 65400, 68100, 71900, None],
        ['Noise-Cancel Headset', 'Europe', 38900, 41200, 43500, 46800, None],
        ['Tablet Stand Adj.', 'Asia Pacific', 12300, 13800, 14900, 16200, None],
        ['Portable SSD 2TB', 'North America', 53200, 56800, 59400, 63100, None],
        ['Ergonomic Chair Pro', 'Europe', 78400, 82100, 85700, 89300, None],
        ['LED Desk Lamp', 'Asia Pacific', 8900, 9400, 10200, 11300, None],
    ]

    for r, row_data in enumerate(sales_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_sales.cell(row=r, column=c, value=val)

    # Add total formulas in column G
    for r in range(2, 14):
        ws_sales.cell(row=r, column=7, value=f'=SUM(C{r}:F{r})')

    # Column widths
    ws_sales.column_dimensions['A'].width = 24
    ws_sales.column_dimensions['B'].width = 16
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws_sales.column_dimensions[col_letter].width = 14

    # Number format for revenue columns
    for r in range(2, 14):
        for c in range(3, 8):
            ws_sales.cell(row=r, column=c).number_format = '#,##0'

    # --- Sheet 2: Old Data (to be moved to Archive_2023) ---
    ws_old = wb.create_sheet('Old Data')

    old_headers = ['Employee', 'Department', 'Hire Date', 'Base Salary', 'Bonus', 'Performance Rating']
    for col, h in enumerate(old_headers, 1):
        cell = ws_old.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri')

    old_data = [
        ['Sarah Chen', 'Engineering', '2021-03-15', 92000, 8500, 'Exceeds'],
        ['Marcus Johnson', 'Marketing', '2020-07-01', 78000, 6200, 'Meets'],
        ['Priya Patel', 'Data Science', '2021-11-20', 105000, 12000, 'Exceeds'],
        ['James Wilson', 'Sales', '2019-05-10', 68000, 15300, 'Exceeds'],
        ['Elena Rodriguez', 'Engineering', '2022-01-08', 88000, 7100, 'Meets'],
        ['David Kim', 'Product', '2020-09-14', 95000, 9800, 'Exceeds'],
        ['Aisha Mohammed', 'HR', '2021-06-22', 72000, 5400, 'Meets'],
        ['Robert Taylor', 'Finance', '2019-12-03', 85000, 7800, 'Meets'],
        ['Lisa Wang', 'Engineering', '2022-04-17', 97000, 10500, 'Exceeds'],
        ['Michael O\'Brien', 'Sales', '2020-02-28', 71000, 13200, 'Exceeds'],
        ['Fatima Al-Hassan', 'Data Science', '2021-08-09', 101000, 11000, 'Meets'],
        ['Carlos Mendez', 'Marketing', '2022-10-15', 74000, 5800, 'Needs Improvement'],
    ]

    for r, row_data in enumerate(old_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_old.cell(row=r, column=c, value=val)

    # Column widths for Old Data
    ws_old.column_dimensions['A'].width = 20
    ws_old.column_dimensions['B'].width = 16
    ws_old.column_dimensions['C'].width = 14
    ws_old.column_dimensions['D'].width = 14
    ws_old.column_dimensions['E'].width = 12
    ws_old.column_dimensions['F'].width = 20

    # Format salary and bonus columns
    for r in range(2, 14):
        ws_old.cell(row=r, column=4).number_format = '#,##0'
        ws_old.cell(row=r, column=5).number_format = '#,##0'

    # --- Sheet 3: Summary ---
    ws_summary = wb.create_sheet('Summary')

    ws_summary.cell(row=1, column=1, value='Department Summary')
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14, name='Calibri')

    summary_headers = ['Department', 'Headcount', 'Avg Salary', 'Budget Allocated']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri')

    summary_data = [
        ['Engineering', 42, 94500, 4200000],
        ['Marketing', 28, 76000, 2350000],
        ['Data Science', 15, 103000, 1680000],
        ['Sales', 35, 69500, 2650000],
        ['Product', 18, 91000, 1780000],
        ['HR', 12, 71000, 920000],
        ['Finance', 20, 83000, 1780000],
    ]

    for r, row_data in enumerate(summary_data, 4):
        for c, val in enumerate(row_data, 1):
            ws_summary.cell(row=r, column=c, value=val)

    for r in range(4, 11):
        ws_summary.cell(row=r, column=3).number_format = '#,##0'
        ws_summary.cell(row=r, column=4).number_format = '#,##0'

    ws_summary.column_dimensions['A'].width = 16
    ws_summary.column_dimensions['B'].width = 12
    ws_summary.column_dimensions['C'].width = 14
    ws_summary.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
