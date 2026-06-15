"""
Initial Setup: Set page to A4 landscape with margins, insert watermark, set print order
Task ID: calc_gg3_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

# Install dependencies on VM
subprocess.run(['pip3', 'install', 'openpyxl', 'Pillow'], capture_output=True)

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
WATERMARK_PATH = '/home/user/Desktop/watermark.png'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_watermark_image():
    """Create a light gray CONFIDENTIAL watermark PNG at /root/Desktop/watermark.png."""
    from PIL import Image, ImageDraw, ImageFont

    width, height = 800, 600
    img = Image.new('RGBA', (width, height), (255, 255, 255, 0))
    draw = ImageDraw.Draw(img)

    # Draw "CONFIDENTIAL" text diagonally in light gray
    text = "CONFIDENTIAL"
    # Try to use a large font
    try:
        font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf", 72)
    except (IOError, OSError):
        try:
            font = ImageFont.truetype("/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf", 72)
        except (IOError, OSError):
            font = ImageFont.load_default()

    # Draw the text multiple times across the image for watermark effect
    text_color = (200, 200, 200, 128)  # light gray, semi-transparent

    # Create a temporary image to rotate
    txt_img = Image.new('RGBA', (width * 2, height * 2), (255, 255, 255, 0))
    txt_draw = ImageDraw.Draw(txt_img)

    # Place text at multiple positions
    y_positions = range(50, height * 2, 200)
    for y in y_positions:
        for x in range(-200, width * 2, 500):
            txt_draw.text((x, y), text, font=font, fill=text_color)

    # Rotate 30 degrees
    txt_img = txt_img.rotate(30, expand=False, center=(width, height))

    # Crop to original size
    left = (txt_img.width - width) // 2
    top = (txt_img.height - height) // 2
    txt_img = txt_img.crop((left, top, left + width, top + height))

    # Save
    os.makedirs(os.path.dirname(WATERMARK_PATH), exist_ok=True)
    txt_img.save(WATERMARK_PATH, 'PNG')
    print(f'Watermark image created: {WATERMARK_PATH}')


def create_initial():
    import openpyxl

    wb = openpyxl.Workbook()

    # --- Print sheet with data in A1:H45 ---
    ws = wb.active
    ws.title = 'Print'

    # Headers (row 1)
    headers = [
        'Employee ID', 'Full Name', 'Department', 'Position',
        'Hire Date', 'Base Salary', 'Bonus', 'Total Compensation'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data (rows 2-45 = 44 rows of data)
    departments = ['Engineering', 'Marketing', 'Finance', 'Human Resources', 'Operations', 'Sales', 'Legal', 'IT Support']
    positions = {
        'Engineering': ['Software Engineer', 'Senior Developer', 'Tech Lead', 'DevOps Engineer', 'QA Analyst'],
        'Marketing': ['Marketing Analyst', 'Content Strategist', 'Brand Manager', 'SEO Specialist', 'Campaign Director'],
        'Finance': ['Financial Analyst', 'Accountant', 'Budget Manager', 'Auditor', 'Controller'],
        'Human Resources': ['HR Coordinator', 'Recruiter', 'Benefits Analyst', 'Training Manager', 'HR Director'],
        'Operations': ['Operations Manager', 'Supply Chain Analyst', 'Logistics Coordinator', 'Process Engineer', 'Facilities Manager'],
        'Sales': ['Sales Representative', 'Account Manager', 'Sales Director', 'Business Development', 'Regional Manager'],
        'Legal': ['Legal Counsel', 'Compliance Officer', 'Paralegal', 'Contract Specialist', 'General Counsel'],
        'IT Support': ['Help Desk Tech', 'System Administrator', 'Network Engineer', 'Security Analyst', 'IT Manager'],
    }

    employees = [
        ('E1001', 'Sarah Chen', 'Engineering', 'Senior Developer', '2021-03-15', 112000, 15000),
        ('E1002', 'Marcus Johnson', 'Marketing', 'Brand Manager', '2020-07-01', 89000, 10000),
        ('E1003', 'Priya Patel', 'Finance', 'Financial Analyst', '2022-01-10', 78000, 8500),
        ('E1004', 'James Williams', 'Engineering', 'Tech Lead', '2019-11-20', 135000, 20000),
        ('E1005', 'Aisha Rahman', 'Human Resources', 'HR Director', '2018-05-08', 105000, 14000),
        ('E1006', 'David Kim', 'Operations', 'Operations Manager', '2021-09-15', 92000, 11500),
        ('E1007', 'Elena Rodriguez', 'Sales', 'Account Manager', '2022-06-01', 76000, 18000),
        ('E1008', 'Thomas Brown', 'Legal', 'Compliance Officer', '2020-02-14', 98000, 12000),
        ('E1009', 'Lisa Wang', 'IT Support', 'System Administrator', '2021-08-22', 85000, 9000),
        ('E1010', 'Robert Martinez', 'Engineering', 'Software Engineer', '2023-01-09', 95000, 10500),
        ('E1011', 'Jennifer Lee', 'Marketing', 'Content Strategist', '2022-04-18', 72000, 7500),
        ('E1012', 'Michael Thompson', 'Finance', 'Controller', '2017-10-30', 125000, 18000),
        ('E1013', 'Amanda Foster', 'Human Resources', 'Recruiter', '2023-03-05', 65000, 6000),
        ('E1014', 'Christopher Davis', 'Operations', 'Supply Chain Analyst', '2021-12-01', 74000, 8000),
        ('E1015', 'Sophia Nguyen', 'Sales', 'Sales Director', '2019-06-15', 115000, 25000),
        ('E1016', 'Daniel Harris', 'Legal', 'Legal Counsel', '2020-09-10', 130000, 16000),
        ('E1017', 'Rachel Green', 'IT Support', 'Network Engineer', '2022-07-22', 88000, 9500),
        ('E1018', 'Kevin O\'Brien', 'Engineering', 'DevOps Engineer', '2021-05-12', 105000, 13000),
        ('E1019', 'Maria Garcia', 'Marketing', 'Campaign Director', '2019-08-28', 98000, 14000),
        ('E1020', 'Andrew Wilson', 'Finance', 'Auditor', '2023-02-14', 82000, 8000),
        ('E1021', 'Nicole Taylor', 'Human Resources', 'Benefits Analyst', '2022-10-01', 68000, 7000),
        ('E1022', 'Brandon Clark', 'Operations', 'Process Engineer', '2020-11-15', 91000, 11000),
        ('E1023', 'Stephanie Adams', 'Sales', 'Business Development', '2021-04-07', 83000, 16000),
        ('E1024', 'Patrick Murphy', 'Legal', 'Paralegal', '2023-06-20', 58000, 5000),
        ('E1025', 'Catherine Zhao', 'IT Support', 'Security Analyst', '2022-01-30', 96000, 12000),
        ('E1026', 'Ryan Jackson', 'Engineering', 'QA Analyst', '2022-09-12', 79000, 8500),
        ('E1027', 'Michelle White', 'Marketing', 'SEO Specialist', '2023-05-15', 68000, 7000),
        ('E1028', 'Jonathan Moore', 'Finance', 'Budget Manager', '2020-03-22', 95000, 12000),
        ('E1029', 'Laura Scott', 'Human Resources', 'Training Manager', '2021-07-18', 78000, 9000),
        ('E1030', 'Eric Turner', 'Operations', 'Logistics Coordinator', '2023-08-01', 62000, 6500),
        ('E1031', 'Diana Cruz', 'Sales', 'Regional Manager', '2018-12-10', 108000, 22000),
        ('E1032', 'Samuel Reed', 'Legal', 'Contract Specialist', '2022-05-25', 75000, 8000),
        ('E1033', 'Angela Brooks', 'IT Support', 'Help Desk Tech', '2023-09-15', 52000, 4500),
        ('E1034', 'Tyler Bennett', 'Engineering', 'Software Engineer', '2023-04-01', 92000, 10000),
        ('E1035', 'Victoria Hayes', 'Marketing', 'Marketing Analyst', '2022-11-08', 70000, 7500),
        ('E1036', 'Mark Sullivan', 'Finance', 'Accountant', '2021-06-14', 72000, 7500),
        ('E1037', 'Natalie Cooper', 'Human Resources', 'HR Coordinator', '2023-07-22', 58000, 5500),
        ('E1038', 'Gregory Price', 'Operations', 'Facilities Manager', '2020-08-30', 80000, 9000),
        ('E1039', 'Jessica Howard', 'Sales', 'Sales Representative', '2023-10-01', 60000, 12000),
        ('E1040', 'William Bell', 'Legal', 'General Counsel', '2016-04-15', 165000, 25000),
        ('E1041', 'Olivia Peterson', 'IT Support', 'IT Manager', '2019-02-28', 110000, 14000),
        ('E1042', 'Alexander Ross', 'Engineering', 'Senior Developer', '2020-10-05', 118000, 16000),
        ('E1043', 'Hannah Morgan', 'Marketing', 'Brand Manager', '2021-01-20', 86000, 10000),
        ('E1044', 'Benjamin Ward', 'Finance', 'Financial Analyst', '2023-11-12', 76000, 8000),
    ]

    for r, emp in enumerate(employees, 2):
        emp_id, name, dept, pos, hire_date, salary, bonus = emp
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=dept)
        ws.cell(row=r, column=4, value=pos)
        ws.cell(row=r, column=5, value=hire_date)
        ws.cell(row=r, column=6, value=salary)
        ws.cell(row=r, column=7, value=bonus)
        ws.cell(row=r, column=8, value=salary + bonus)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


# Create watermark image first, then spreadsheet
create_watermark_image()
create_initial()
