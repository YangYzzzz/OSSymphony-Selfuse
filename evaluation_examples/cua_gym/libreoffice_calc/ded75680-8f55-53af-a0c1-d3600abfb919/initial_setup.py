"""
Initial Setup: Q2 Expense Category Pie Chart Task
Task ID: calc_fin_expense_category_pie_018
Domain: libreoffice_calc

Creates:
  - Sheet 'RawExpenses': 119 rows of individual expense entries (rows 2-120)
    with Date, Category, Description, Amount
  - Sheet 'Summary': category list in A2:A8, headers in row 1, B/C columns empty
    (NO formulas, NO chart, NO formatting)
"""

import os
from datetime import date, timedelta
import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_expense_category_pie_018'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: RawExpenses                                                 #
    # ------------------------------------------------------------------ #
    ws_raw = wb.active
    ws_raw.title = 'RawExpenses'

    # Headers
    ws_raw['A1'] = 'Date'
    ws_raw['B1'] = 'Category'
    ws_raw['C1'] = 'Description'
    ws_raw['D1'] = 'Amount'

    categories = [
        'Travel', 'Meals', 'Software', 'Equipment', 'Marketing', 'Consulting', 'Facilities'
    ]

    descriptions = {
        'Travel': [
            'Flight to NYC – Q2 conference',
            'Hotel stay – client meeting Chicago',
            'Rental car – Boston product demo',
            'Train ticket – DC summit',
            'Taxi/rideshare – airport transfers',
            'Parking fees – downtown office',
            'Tolls – cross-state delivery run',
        ],
        'Meals': [
            'Team lunch – project kickoff',
            'Client dinner – contract signing',
            'Coffee & snacks – all-hands meeting',
            'Working lunch – budget review',
            'Catering – product launch event',
            'Team dinner – end-of-sprint celebration',
            'Breakfast – early investor call',
        ],
        'Software': [
            'Adobe Creative Cloud – annual renewal',
            'Slack Business+ – monthly subscription',
            'Zoom Pro – video conferencing',
            'GitHub Enterprise – dev licenses',
            'Jira Software – project management',
            'Figma Organization – design tools',
            'AWS monthly usage bill',
        ],
        'Equipment': [
            'Standing desk – remote employee setup',
            'MacBook Pro – new hire onboarding',
            'External monitors x2 – workstation upgrade',
            'Wireless keyboard & mouse – office refresh',
            'USB-C hub – laptop accessories',
            'Webcam HD – video conferencing upgrade',
            'Office chair – ergonomic replacement',
        ],
        'Marketing': [
            'Google Ads – Q2 campaign',
            'LinkedIn Ads – recruiter pipeline',
            'Sponsored blog post – industry outlet',
            'Trade show booth – TechWorld Expo',
            'Brochure printing – sales collateral',
            'Email platform – Mailchimp subscription',
            'Social media management tool',
        ],
        'Consulting': [
            'Legal advice – contract review',
            'Accounting firm – quarterly audit',
            'HR consultant – policy update',
            'IT security assessment',
            'UX research firm – usability study',
            'Market research – competitive analysis',
            'Strategy consultant – growth planning',
        ],
        'Facilities': [
            'Office rent – April',
            'Office rent – May',
            'Office rent – June',
            'Electricity bill – Q2',
            'Internet & phone – monthly service',
            'Cleaning service – bi-weekly',
            'Building maintenance – HVAC repair',
        ],
    }

    amount_ranges = {
        'Travel':      (150, 2800),
        'Meals':       (25, 450),
        'Software':    (20, 1200),
        'Equipment':   (80, 3500),
        'Marketing':   (200, 5000),
        'Consulting':  (500, 8000),
        'Facilities':  (800, 6000),
    }

    start_date = date(2025, 4, 1)
    end_date   = date(2025, 6, 30)
    date_range = (end_date - start_date).days

    rows = []
    # Ensure at least a handful of entries per category
    for cat in categories:
        for desc in descriptions[cat]:
            days_offset = random.randint(0, date_range)
            entry_date = start_date + timedelta(days=days_offset)
            low, high = amount_ranges[cat]
            amount = round(random.uniform(low, high), 2)
            rows.append((entry_date.strftime('%Y-%m-%d'), cat, desc, amount))

    # Fill up to 119 rows with more random entries
    extra_cats_weights = [0.14, 0.10, 0.12, 0.12, 0.16, 0.20, 0.16]
    while len(rows) < 119:
        cat = random.choices(categories, weights=extra_cats_weights, k=1)[0]
        desc = random.choice(descriptions[cat])
        days_offset = random.randint(0, date_range)
        entry_date = start_date + timedelta(days=days_offset)
        low, high = amount_ranges[cat]
        amount = round(random.uniform(low, high), 2)
        rows.append((entry_date.strftime('%Y-%m-%d'), cat, desc, amount))

    # Sort by date
    rows.sort(key=lambda x: x[0])

    for r, (d, cat, desc, amount) in enumerate(rows, 2):
        ws_raw.cell(row=r, column=1, value=d)
        ws_raw.cell(row=r, column=2, value=cat)
        ws_raw.cell(row=r, column=3, value=desc)
        ws_raw.cell(row=r, column=4, value=amount)

    # Column widths for readability
    ws_raw.column_dimensions['A'].width = 14
    ws_raw.column_dimensions['B'].width = 16
    ws_raw.column_dimensions['C'].width = 42
    ws_raw.column_dimensions['D'].width = 12

    # ------------------------------------------------------------------ #
    # Sheet 2: Summary (category list only, NO formulas, NO chart)        #
    # ------------------------------------------------------------------ #
    ws_sum = wb.create_sheet('Summary')

    ws_sum['A1'] = 'Category'
    ws_sum['B1'] = 'Total'
    ws_sum['C1'] = 'Pct of Total'

    for i, cat in enumerate(categories, 2):
        ws_sum.cell(row=i, column=1, value=cat)
    # B2:C8 intentionally empty — agent must fill these in

    ws_sum.column_dimensions['A'].width = 16
    ws_sum.column_dimensions['B'].width = 14
    ws_sum.column_dimensions['C'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  RawExpenses: {ws_raw.max_row - 1} data rows')
    print(f'  Summary: categories in A2:A8, B/C columns empty')

create_initial()
