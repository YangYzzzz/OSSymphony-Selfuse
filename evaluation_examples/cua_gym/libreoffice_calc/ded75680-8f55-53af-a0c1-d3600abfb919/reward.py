"""
Reward Script: Q2 Expense Category Pie Chart with SUMIF formulas
Task ID: calc_fin_expense_category_pie_018
Domain: libreoffice_calc

Scoring Rubric (total = 1.0):
  Component 1: SUMIF formulas in B2:B8 on Summary sheet            — 0.35 pts
  Component 2: Grand Total row (A9='Grand Total' bold, B9=SUM)     — 0.20 pts
  Component 3: Percentage formulas in C2:C8 with % format           — 0.20 pts
  Component 4: Pie chart on Summary sheet with correct title        — 0.15 pts
  Component 5: Chart data references B2:B8 and has % data labels   — 0.10 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_expense_category_pie_018'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: SUMIF formulas in B2:B8 (0.35 points)
    # Initial: B2:B8 are all None
    # Golden: each cell has =SUMIF(RawExpenses.$B$2:$B$120,An,RawExpenses.$D$2:$D$120)
    try:
        sumif_count = 0
        sumif_with_currency = 0
        for row in range(2, 9):
            cell_b = ws.cell(row=row, column=2)
            val = cell_b.value
            if val and isinstance(val, str) and 'SUMIF' in val.upper():
                sumif_count += 1
                # Check if currency format is also applied
                fmt = cell_b.number_format
                if fmt and ('$' in fmt or '#,##0' in fmt):
                    sumif_with_currency += 1

        if sumif_count == 7 and sumif_with_currency == 7:
            print(f"PASS: Component 1 — All 7 SUMIF formulas in B2:B8 with currency format (0.35 pts)")
            total_score += 0.35
        elif sumif_count == 7:
            print(f"PASS (partial): Component 1 — All 7 SUMIF formulas found but only {sumif_with_currency}/7 have currency format (0.25 pts)")
            total_score += 0.25
        elif sumif_count > 0:
            partial = round(0.35 * sumif_count / 7, 2)
            print(f"PARTIAL: Component 1 — Only {sumif_count}/7 SUMIF formulas in B2:B8 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No SUMIF formulas found in B2:B8 (expected 7, found {sumif_count})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Grand Total row — A9='Grand Total' (bold), B9=SUM formula with currency format (0.20 points)
    # Initial: A9=None, B9=None
    # Golden: A9='Grand Total', bold=True; B9='=SUM(B2:B8)', bold=True, fmt=$#,##0.00
    try:
        a9 = ws.cell(row=9, column=1)
        b9 = ws.cell(row=9, column=2)

        a9_ok = (a9.value == 'Grand Total')
        a9_bold = (a9.font.bold == True)
        b9_formula = (b9.value and isinstance(b9.value, str) and 'SUM' in b9.value.upper() and 'B2' in b9.value and 'B8' in b9.value)
        b9_bold = (b9.font.bold == True)
        b9_currency = (b9.number_format and ('$' in b9.number_format or '#,##0' in b9.number_format))

        if a9_ok and a9_bold and b9_formula and b9_bold and b9_currency:
            print(f"PASS: Component 2 — Grand Total row: A9='{a9.value}' (bold={a9_bold}), B9='{b9.value}' (bold={b9_bold}, fmt={b9.number_format}) (0.20 pts)")
            total_score += 0.20
        elif a9_ok and b9_formula:
            print(f"PARTIAL: Component 2 — Grand Total row found but missing bold/format. A9_bold={a9_bold}, B9_bold={b9_bold}, B9_currency={b9_currency} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Grand Total row missing. A9={repr(a9.value)}, B9={repr(b9.value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage formulas in C2:C8 with % number format (0.20 points)
    # Initial: C2:C8 are all None
    # Golden: each cell has =Bn/$B$9 with number_format containing '%'
    try:
        pct_formula_count = 0
        pct_format_count = 0
        for row in range(2, 9):
            cell_c = ws.cell(row=row, column=3)
            val = cell_c.value
            has_formula = (val and isinstance(val, str) and '$B$9' in val and 'B' in val)
            has_pct_fmt = (cell_c.number_format and '%' in cell_c.number_format)
            if has_formula:
                pct_formula_count += 1
            if has_formula and has_pct_fmt:
                pct_format_count += 1

        if pct_format_count == 7:
            print(f"PASS: Component 3 — All 7 percentage formulas in C2:C8 with % format (0.20 pts)")
            total_score += 0.20
        elif pct_formula_count == 7:
            print(f"PARTIAL: Component 3 — All 7 % formulas found but only {pct_format_count}/7 have % format (0.12 pts)")
            total_score += 0.12
        elif pct_formula_count > 0:
            partial = round(0.20 * pct_formula_count / 7, 2)
            print(f"PARTIAL: Component 3 — Only {pct_formula_count}/7 percentage formulas in C2:C8 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No percentage formulas found in C2:C8 (expected 7, found {pct_formula_count})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Pie chart present on Summary sheet with title 'Q2 Expense Breakdown by Category' (0.15 points)
    # Initial: no charts
    # Golden: 1 PieChart with correct title
    try:
        charts = ws._charts
        pie_charts = [c for c in charts if type(c).__name__ == 'PieChart']

        if not pie_charts:
            print(f"FAIL: Component 4 — No pie chart found on Summary sheet (found {len(charts)} charts total)")
        else:
            chart = pie_charts[0]

            # Extract title text
            chart_title = None
            try:
                chart_title = chart.title.tx.rich.p[0].r[0].t
            except Exception:
                try:
                    chart_title = str(chart.title)
                except Exception:
                    chart_title = None

            expected_title = 'Q2 Expense Breakdown by Category'
            title_ok = (chart_title and expected_title.lower() in chart_title.lower())

            if title_ok:
                print(f"PASS: Component 4 — Pie chart found with correct title '{chart_title}' (0.15 pts)")
                total_score += 0.15
            elif not title_ok and chart_title is not None:
                print(f"PARTIAL: Component 4 — Pie chart found but title mismatch: expected '{expected_title}', got '{chart_title}' (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 4 — Pie chart found but no title detected")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Chart data references B2:B8 (values) and A2:A8 (categories), has % data labels (0.10 points)
    # Initial: no charts
    # Golden: series val ref = 'Summary'!$B$2:$B$8, cat ref = 'Summary'!$A$2:$A$8, showPercent=True
    try:
        charts = ws._charts
        pie_charts = [c for c in charts if type(c).__name__ == 'PieChart']

        if not pie_charts:
            print(f"FAIL: Component 5 — No pie chart to check references")
        else:
            chart = pie_charts[0]
            series = chart.series

            if not series:
                print(f"FAIL: Component 5 — Pie chart has no series")
            else:
                ser = series[0]

                # Check value reference
                # Note: openpyxl stores references as "'Summary'!$B$2:$B$8" (with sheet name in single quotes)
                val_ref_ok = False
                try:
                    val_ref = ser.val.numRef.f  # e.g. "'Summary'!$B$2:$B$8"
                    # Normalize: strip quotes and check key parts
                    val_ref_norm = val_ref.replace("'", "").replace("$", "").upper()
                    val_ref_ok = ('B2' in val_ref_norm and 'B8' in val_ref_norm and 'SUMMARY' in val_ref_norm)
                except Exception:
                    pass

                # Check category reference
                cat_ref_ok = False
                try:
                    # Cat could be in numRef or strRef
                    cat_ref = None
                    try:
                        cat_ref = ser.cat.numRef.f
                    except Exception:
                        pass
                    if cat_ref is None:
                        try:
                            cat_ref = ser.cat.strRef.f
                        except Exception:
                            pass
                    if cat_ref:
                        cat_ref_norm = cat_ref.replace("'", "").replace("$", "").upper()
                        cat_ref_ok = ('A2' in cat_ref_norm and 'A8' in cat_ref_norm and 'SUMMARY' in cat_ref_norm)
                except Exception:
                    pass

                # Check data labels show percentage
                show_pct = False
                try:
                    show_pct = (chart.dLbls is not None and chart.dLbls.showPercent == True)
                except Exception:
                    pass

                passed_checks = sum([val_ref_ok, cat_ref_ok, show_pct])

                if passed_checks == 3:
                    print(f"PASS: Component 5 — Chart refs B2:B8/A2:A8, showPercent=True (0.10 pts)")
                    total_score += 0.10
                elif passed_checks >= 2:
                    print(f"PARTIAL: Component 5 — {passed_checks}/3 checks passed: val_ref={val_ref_ok}, cat_ref={cat_ref_ok}, show_pct={show_pct} (0.06 pts)")
                    total_score += 0.06
                elif passed_checks == 1:
                    print(f"PARTIAL: Component 5 — Only {passed_checks}/3 checks passed: val_ref={val_ref_ok}, cat_ref={cat_ref_ok}, show_pct={show_pct} (0.03 pts)")
                    total_score += 0.03
                else:
                    print(f"FAIL: Component 5 — All checks failed: val_ref={val_ref_ok}, cat_ref={cat_ref_ok}, show_pct={show_pct}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
