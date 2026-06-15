"""
Reward Script: Track consignment inventory with quantity remaining, value remaining,
days since movement, stale flags, conditional formatting, and summary section.
Task ID: calc_ops_inventory_consignment_tracking_060
Domain: libreoffice_calc
Scoring:
  - Component 1: Qty Remaining formulas F2:F61 (=D-E)           0.25 pts
  - Component 2: Value Remaining formulas H2:H61 (=F*G)         0.20 pts
  - Component 3: Days Since Movement formulas J2:J61 (=TODAY()-I) 0.15 pts
  - Component 4: Stale Flag formulas K2:K61 (=IF(AND(J>60,F>0),"STALE","")) 0.15 pts
  - Component 5: Conditional formatting on K2:K61 (red fill)    0.10 pts
  - Component 6: Summary section with SUMIF by customer + grand total 0.15 pts
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_consignment_tracking_060'

SHEET_NAME = 'ConsignmentStock'
DATA_ROWS = range(2, 62)   # rows 2-61 (60 records)


def normalize_formula(formula):
    """Normalize formula string for comparison: uppercase, no spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -----------------------------------------------------------------------
    # Component 1: Qty Remaining formulas in F2:F61  (0.25 points)
    # Each cell should contain a formula matching =D{n}-E{n}
    # This FAILS on initial (cells are empty) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        formula_count = 0
        correct_count = 0
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=6).value  # column F
            if val is not None:
                formula_count += 1
                # Accept =D{row}-E{row} pattern
                expected = f'=D{row}-E{row}'.upper()
                if normalize_formula(str(val)) == expected:
                    correct_count += 1

        if correct_count == 60:
            print(f"PASS: Component 1 — All 60 Qty Remaining formulas in F2:F61 are correct (0.25 pts)")
            total_score += 0.25
        elif correct_count >= 50:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {correct_count}/60 Qty Remaining formulas correct ({partial} pts)")
            total_score += partial
        elif correct_count > 0:
            partial = 0.05
            print(f"PARTIAL: Component 1 — {correct_count}/60 Qty Remaining formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected =D{{n}}-E{{n}} formulas in F2:F61, found {formula_count} formula(s), {correct_count} correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Value Remaining formulas in H2:H61  (0.20 points)
    # Each cell should contain a formula matching =F{n}*G{n}
    # This FAILS on initial (cells are empty) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        correct_count = 0
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=8).value  # column H
            if val is not None:
                # Accept =F{row}*G{row} pattern
                expected = f'=F{row}*G{row}'.upper()
                if normalize_formula(str(val)) == expected:
                    correct_count += 1

        if correct_count == 60:
            print(f"PASS: Component 2 — All 60 Value Remaining formulas in H2:H61 are correct (0.20 pts)")
            total_score += 0.20
        elif correct_count >= 50:
            partial = 0.12
            print(f"PARTIAL: Component 2 — {correct_count}/60 Value Remaining formulas correct ({partial} pts)")
            total_score += partial
        elif correct_count > 0:
            partial = 0.05
            print(f"PARTIAL: Component 2 — {correct_count}/60 Value Remaining formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected =F{{n}}*G{{n}} formulas in H2:H61, found {correct_count} correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Days Since Movement formulas in J2:J61  (0.15 points)
    # Each cell should contain =TODAY()-I{n}
    # This FAILS on initial (cells are empty) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        correct_count = 0
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=10).value  # column J
            if val is not None:
                expected = f'=TODAY()-I{row}'.upper()
                if normalize_formula(str(val)) == expected:
                    correct_count += 1

        if correct_count == 60:
            print(f"PASS: Component 3 — All 60 Days Since Movement formulas in J2:J61 are correct (0.15 pts)")
            total_score += 0.15
        elif correct_count >= 50:
            partial = 0.09
            print(f"PARTIAL: Component 3 — {correct_count}/60 Days Since Movement formulas correct ({partial} pts)")
            total_score += partial
        elif correct_count > 0:
            partial = 0.04
            print(f"PARTIAL: Component 3 — {correct_count}/60 Days Since Movement formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected =TODAY()-I{{n}} formulas in J2:J61, found {correct_count} correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Stale Flag formulas in K2:K61  (0.15 points)
    # Each cell should contain =IF(AND(J{n}>60,F{n}>0),"STALE","")
    # This FAILS on initial (cells are empty) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        correct_count = 0
        for row in DATA_ROWS:
            val = ws.cell(row=row, column=11).value  # column K
            if val is not None:
                # Normalize and check pattern
                norm = normalize_formula(str(val))
                # Accept variations: =IF(AND(J{n}>60,F{n}>0),"STALE","")
                expected = f'=IF(AND(J{row}>60,F{row}>0),"STALE","")'.upper()
                if norm == expected:
                    correct_count += 1

        if correct_count == 60:
            print(f"PASS: Component 4 — All 60 Stale Flag formulas in K2:K61 are correct (0.15 pts)")
            total_score += 0.15
        elif correct_count >= 50:
            partial = 0.09
            print(f"PARTIAL: Component 4 — {correct_count}/60 Stale Flag formulas correct ({partial} pts)")
            total_score += partial
        elif correct_count > 0:
            partial = 0.04
            print(f"PARTIAL: Component 4 — {correct_count}/60 Stale Flag formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Expected =IF(AND(J{{n}}>60,F{{n}}>0),\"STALE\",\"\") in K2:K61, found {correct_count} correct")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Conditional formatting on K2:K61 with red fill  (0.10 points)
    # Range K2:K61 should have a conditional formatting rule that fills red when "STALE"
    # This FAILS on initial (no CF applied) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        cf_found = False
        red_fill_found = False
        stale_condition_found = False

        for cf_range, rules in ws.conditional_formatting._cf_rules.items():
            range_str = str(cf_range)
            # Check if the CF range covers K2:K61
            if 'K2' in range_str and 'K61' in range_str:
                cf_found = True
                for rule in rules:
                    # Check for formula mentioning STALE or matching K expression
                    if rule.formula:
                        formula_str = ' '.join(str(f) for f in rule.formula).upper()
                        if 'STALE' in formula_str or 'K2' in formula_str:
                            stale_condition_found = True
                    # Check for red fill
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            # Check for red color variants: FFFF0000, FF0000, etc.
                            if 'FF0000' in fill_color.upper():
                                red_fill_found = True
                        except Exception:
                            pass

        if cf_found and stale_condition_found and red_fill_found:
            print(f"PASS: Component 5 — Conditional formatting on K2:K61 with red fill for STALE (0.10 pts)")
            total_score += 0.10
        elif cf_found and (stale_condition_found or red_fill_found):
            partial = 0.05
            print(f"PARTIAL: Component 5 — CF found on K range but incomplete (stale_cond={stale_condition_found}, red_fill={red_fill_found}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No conditional formatting found on K2:K61 with red fill (cf_found={cf_found})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -----------------------------------------------------------------------
    # Component 6: Summary section with SUMIF by customer and GRAND TOTAL  (0.15 points)
    # Should have label at A64/A65, SUMIF formulas in B66:B75 for 10 customers,
    # and GRAND TOTAL at A76 with =SUM(H2:H61) or =SUM($H$2:$H$61) in B76
    # This FAILS on initial (no summary section) -> PASSES on golden
    # -----------------------------------------------------------------------
    try:
        # Check for summary header label
        header_present = False
        label_cell = ws.cell(row=64, column=1).value
        if label_cell and 'CONSIGNMENT' in str(label_cell).upper():
            header_present = True
        # Also try row 63
        if not header_present:
            label_cell = ws.cell(row=63, column=1).value
            if label_cell and 'CONSIGNMENT' in str(label_cell).upper():
                header_present = True

        # Check for SUMIF formulas in the summary block (search rows 63-85)
        sumif_count = 0
        grand_total_found = False
        for row in range(60, 90):
            val_a = ws.cell(row=row, column=1).value
            val_b = ws.cell(row=row, column=2).value
            if val_b and isinstance(val_b, str):
                norm = normalize_formula(val_b)
                if 'SUMIF' in norm and 'H$2' in norm.replace(' ', '').upper():
                    sumif_count += 1
                if 'SUM' in norm and 'H' in norm and ('$H$2:$H$61' in norm.upper() or 'H2:H61' in norm.upper()):
                    # Could be SUM or SUMIF
                    if norm.startswith('=SUM(') and 'SUMIF' not in norm:
                        grand_total_found = True
            if val_a and 'GRAND TOTAL' in str(val_a).upper():
                # verify grand total formula in B
                bt = ws.cell(row=row, column=2).value
                if bt and isinstance(bt, str) and 'SUM' in bt.upper():
                    grand_total_found = True

        if sumif_count >= 10 and grand_total_found:
            print(f"PASS: Component 6 — Summary section with {sumif_count} SUMIF formulas and grand total found (0.15 pts)")
            total_score += 0.15
        elif sumif_count >= 5 and grand_total_found:
            partial = 0.10
            print(f"PARTIAL: Component 6 — Summary has {sumif_count} SUMIF formulas and grand total ({partial} pts)")
            total_score += partial
        elif sumif_count >= 3:
            partial = 0.07
            print(f"PARTIAL: Component 6 — Summary has {sumif_count} SUMIF formulas but grand total={grand_total_found} ({partial} pts)")
            total_score += partial
        elif grand_total_found:
            partial = 0.05
            print(f"PARTIAL: Component 6 — Grand total found but only {sumif_count} SUMIF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Summary section incomplete: header={header_present}, sumif_count={sumif_count}, grand_total={grand_total_found}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
