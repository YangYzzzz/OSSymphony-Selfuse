"""
Initial Setup: Track consignment inventory held at customer sites
Task ID: calc_ops_inventory_consignment_tracking_060
Domain: libreoffice_calc

Creates ConsignmentStock sheet with:
- 60 consignment records (rows 2-61)
- Headers: Customer, SKU, Product, Qty Sent, Qty Consumed, Qty Remaining (empty),
           Unit Cost, Value Remaining (empty), Last Movement Date,
           Days Since Movement (empty), Stale Flag (empty)
- Columns F, H, J, K are intentionally left empty (to be filled by agent)
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_inventory_consignment_tracking_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ConsignmentStock'

    # --- Header row ---
    headers = [
        'Customer', 'SKU', 'Product', 'Qty Sent', 'Qty Consumed',
        'Qty Remaining', 'Unit Cost', 'Value Remaining',
        'Last Movement Date', 'Days Since Movement', 'Stale Flag'
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[1].height = 30

    # --- Realistic data ---
    customers = [
        'Apex Manufacturing Ltd',
        'Riverstone Distributors',
        'Pacific Retail Group',
        'Summit Industrial Co.',
        'Bluewood Supply Chain',
        'Harmon Logistics Inc.',
        'Crestview Electronics',
        'Northgate Hardware',
        'Sunridge Auto Parts',
        'Lakewood Medical Supplies',
    ]

    products = [
        ('SKU-1001', 'Steel Fasteners M6x20', 12.50),
        ('SKU-1002', 'Hydraulic Seal Kit Type A', 45.00),
        ('SKU-1003', 'Industrial Filter Cartridge', 28.75),
        ('SKU-1004', 'Bearing Assembly 6205', 18.90),
        ('SKU-1005', 'Valve Actuator 24V', 135.00),
        ('SKU-1006', 'Conveyor Belt Section 2m', 220.00),
        ('SKU-1007', 'Lubricant Grease 5kg Drum', 34.20),
        ('SKU-1008', 'Coupling Flange DN50', 67.80),
        ('SKU-1009', 'Safety Gloves Cut Level 5', 9.95),
        ('SKU-1010', 'Electrical Cable 6mm 50m', 89.50),
        ('SKU-1011', 'Pneumatic Cylinder 80mm', 156.40),
        ('SKU-1012', 'LED Panel Light 600x600', 48.00),
        ('SKU-1013', 'Stainless Pipe Elbow 90deg', 22.30),
        ('SKU-1014', 'Motor Drive 5.5kW VFD', 310.00),
        ('SKU-1015', 'Paint Epoxy Primer 20L', 55.60),
    ]

    today = date(2026, 3, 4)

    # 60 records with varied movement dates - some stale (>60 days ago), some recent
    import random
    random.seed(42)

    records = []
    for i in range(60):
        customer = customers[i % len(customers)]
        sku, product, unit_cost = products[i % len(products)]
        # Vary qty sent and consumed
        qty_sent = random.choice([50, 100, 150, 200, 250, 300, 500])
        consumed_pct = random.uniform(0.1, 0.85)
        qty_consumed = int(qty_sent * consumed_pct)
        # Vary movement dates: ~30% stale (61-180 days ago), rest recent (1-59 days)
        if i % 3 == 0:
            days_ago = random.randint(61, 180)
        else:
            days_ago = random.randint(1, 59)
        last_movement = today - timedelta(days=days_ago)

        # Slightly vary unit cost per customer
        adjusted_cost = round(unit_cost * random.uniform(0.95, 1.05), 2)

        records.append((customer, sku, product, qty_sent, qty_consumed, adjusted_cost, last_movement))

    thin = Side(style='thin', color='FFB8CCE4')
    for r, (customer, sku, product, qty_sent, qty_consumed, unit_cost, last_movement) in enumerate(records, 2):
        # A: Customer
        ws.cell(row=r, column=1, value=customer)
        # B: SKU
        ws.cell(row=r, column=2, value=sku)
        # C: Product
        ws.cell(row=r, column=3, value=product)
        # D: Qty Sent
        ws.cell(row=r, column=4, value=qty_sent)
        # E: Qty Consumed
        ws.cell(row=r, column=5, value=qty_consumed)
        # F: Qty Remaining - INTENTIONALLY EMPTY (agent fills)
        # ws.cell(row=r, column=6) - leave empty
        # G: Unit Cost
        cell_g = ws.cell(row=r, column=7, value=unit_cost)
        cell_g.number_format = '$#,##0.00'
        # H: Value Remaining - INTENTIONALLY EMPTY (agent fills)
        # ws.cell(row=r, column=8) - leave empty
        # I: Last Movement Date
        cell_i = ws.cell(row=r, column=9, value=last_movement)
        cell_i.number_format = 'yyyy-mm-dd'
        # J: Days Since Movement - INTENTIONALLY EMPTY (agent fills)
        # ws.cell(row=r, column=10) - leave empty
        # K: Stale Flag - INTENTIONALLY EMPTY (agent fills)
        # ws.cell(row=r, column=11) - leave empty

        # Alternate row shading for readability
        if r % 2 == 0:
            row_fill = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
            for col in range(1, 12):
                ws.cell(row=r, column=col).fill = row_fill

    # --- Column widths ---
    col_widths = {
        'A': 28, 'B': 12, 'C': 30, 'D': 11, 'E': 14,
        'F': 14, 'G': 12, 'H': 16, 'I': 18, 'J': 20, 'K': 12
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Auto-filter on header row
    ws.auto_filter.ref = 'A1:K61'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ConsignmentStock')
    print(f'  Rows: 61 (1 header + 60 data rows)')
    print(f'  Columns A-E and G, I filled; F, H, J, K intentionally empty')


create_initial()
