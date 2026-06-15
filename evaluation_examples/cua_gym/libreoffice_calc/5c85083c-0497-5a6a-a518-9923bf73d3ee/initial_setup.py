"""
Initial Setup: Create a Monthly Template workbook with formulas, formatting, and data validation
Task ID: calc_gsi_059
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Monthly Template sheet ---
    ws = wb.active
    ws.title = 'Monthly Template'

    # Styling definitions
    header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'

    # Title row (merged)
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Monthly Expense Report'
    ws['A1'].font = Font(name='Arial', size=16, bold=True, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    # Subtitle row
    ws.merge_cells('A2:F2')
    ws['A2'] = 'Department: Operations'
    ws['A2'].font = Font(name='Arial', size=10, italic=True, color='666666')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 22

    # Blank row 3
    ws.row_dimensions[3].height = 8

    # Headers in row 4
    headers = ['Category', 'Budget', 'Actual', 'Variance', 'Variance %', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[4].height = 28

    # Column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Data rows (row 5 to 16 = 12 categories)
    categories = [
        ('Personnel Salaries', 45000.00, 44250.00),
        ('Office Supplies', 3500.00, 3875.50),
        ('Travel & Lodging', 8000.00, 9120.00),
        ('Software Licenses', 12500.00, 12500.00),
        ('Equipment Maintenance', 5200.00, 4980.00),
        ('Marketing Materials', 7800.00, 8150.75),
        ('Training & Development', 4500.00, 3200.00),
        ('Utilities', 6200.00, 6385.00),
        ('Insurance Premiums', 9800.00, 9800.00),
        ('Professional Services', 15000.00, 14250.00),
        ('Facility Rental', 22000.00, 22000.00),
        ('Miscellaneous', 2000.00, 1850.00),
    ]

    data_font = Font(name='Arial', size=10)
    cat_font = Font(name='Arial', size=10, bold=True)

    for i, (cat, budget, actual) in enumerate(categories):
        row = 5 + i
        # Category
        c = ws.cell(row=row, column=1, value=cat)
        c.font = cat_font
        c.border = thin_border

        # Budget
        c = ws.cell(row=row, column=2, value=budget)
        c.font = data_font
        c.number_format = currency_fmt
        c.border = thin_border

        # Actual
        c = ws.cell(row=row, column=3, value=actual)
        c.font = data_font
        c.number_format = currency_fmt
        c.border = thin_border

        # Variance formula: Budget - Actual
        c = ws.cell(row=row, column=4, value=f'=B{row}-C{row}')
        c.font = data_font
        c.number_format = currency_fmt
        c.border = thin_border

        # Variance % formula: Variance / Budget
        c = ws.cell(row=row, column=5, value=f'=IF(B{row}=0,0,D{row}/B{row})')
        c.font = data_font
        c.number_format = pct_fmt
        c.border = thin_border

        # Status (empty - populated by data validation)
        c = ws.cell(row=row, column=6)
        c.font = data_font
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    # Totals row (row 17)
    last_data_row = 16
    total_row = 17
    ws.cell(row=total_row, column=1, value='TOTAL').font = Font(name='Arial', size=11, bold=True)
    ws.cell(row=total_row, column=1).border = thin_border

    for col in [2, 3, 4]:
        col_letter = chr(64 + col)  # B, C, D
        c = ws.cell(row=total_row, column=col, value=f'=SUM({col_letter}5:{col_letter}{last_data_row})')
        c.font = Font(name='Arial', size=11, bold=True)
        c.number_format = currency_fmt
        c.border = thin_border

    # Variance % total
    c = ws.cell(row=total_row, column=5, value=f'=IF(B{total_row}=0,0,D{total_row}/B{total_row})')
    c.font = Font(name='Arial', size=11, bold=True)
    c.number_format = pct_fmt
    c.border = thin_border

    ws.cell(row=total_row, column=6).border = thin_border

    # Total row styling
    total_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    for col in range(1, 7):
        ws.cell(row=total_row, column=col).fill = total_fill

    # Data validation for Status column (F5:F16)
    dv = DataValidation(
        type='list',
        formula1='"On Track,Over Budget,Under Budget,Pending"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = 'Please select a valid status'
    dv.errorTitle = 'Invalid Status'
    dv.prompt = 'Select expense status'
    dv.promptTitle = 'Status'
    dv.add(f'F5:F{last_data_row}')
    ws.add_data_validation(dv)

    # Freeze panes at A5 (freeze title + header rows)
    ws.freeze_panes = 'A5'

    # Auto-filter on headers
    ws.auto_filter.ref = f'A4:F{total_row}'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
