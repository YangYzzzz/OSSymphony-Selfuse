"""
Reward Script: Duplicate 'Monthly Template' sheet to create 'March', 'April', 'May' copies
Task ID: calc_gsi_059
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): 'March' sheet exists with content matching template
  Component 2 (0.25): 'April' sheet exists with content matching template
  Component 3 (0.25): 'May' sheet exists with content matching template
  Component 4 (0.15): Correct sheet ordering (copies after template, at end)
  Component 5 (0.10): Original 'Monthly Template' sheet still exists and intact
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_059'


def sheets_content_match(ws_template, ws_copy):
    """Compare cell values between template and a copy sheet."""
    for row in range(1, ws_template.max_row + 1):
        for col in range(1, ws_template.max_column + 1):
            tv = ws_template.cell(row=row, column=col).value
            cv = ws_copy.cell(row=row, column=col).value
            if tv != cv:
                return False, f"({row},{col}): template={tv!r} vs copy={cv!r}"
    return True, "all values match"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Precondition: 'Monthly Template' must exist to compare against
    if 'Monthly Template' not in sheet_names:
        print("CRITICAL: 'Monthly Template' sheet missing — cannot verify copies")
        print("REWARD: 0.0")
        return 0.0

    template = wb['Monthly Template']

    # Component 1: 'March' sheet exists with matching content (0.25 points)
    try:
        if 'March' in sheet_names:
            match, detail = sheets_content_match(template, wb['March'])
            if match:
                print(f"PASS: Component 1 — 'March' sheet exists and content matches template (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — 'March' sheet content mismatch: {detail}")
        else:
            print("FAIL: Component 1 — 'March' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'April' sheet exists with matching content (0.25 points)
    try:
        if 'April' in sheet_names:
            match, detail = sheets_content_match(template, wb['April'])
            if match:
                print(f"PASS: Component 2 — 'April' sheet exists and content matches template (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 2 — 'April' sheet content mismatch: {detail}")
        else:
            print("FAIL: Component 2 — 'April' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 'May' sheet exists with matching content (0.25 points)
    try:
        if 'May' in sheet_names:
            match, detail = sheets_content_match(template, wb['May'])
            if match:
                print(f"PASS: Component 3 — 'May' sheet exists and content matches template (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — 'May' sheet content mismatch: {detail}")
        else:
            print("FAIL: Component 3 — 'May' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Sheet order — copies placed after template at end of workbook (0.15 points)
    # The task says "placing each copy at the end of the workbook"
    # Expected order: 'Monthly Template' first, then 'March', 'April', 'May' at the end
    try:
        required_copies = ['March', 'April', 'May']
        all_present = all(name in sheet_names for name in required_copies)
        if all_present:
            template_idx = sheet_names.index('Monthly Template')
            march_idx = sheet_names.index('March')
            april_idx = sheet_names.index('April')
            may_idx = sheet_names.index('May')
            # Copies must be after template and in correct relative order at end
            correct_order = (template_idx < march_idx < april_idx < may_idx)
            # They should be the last 3 sheets
            at_end = (may_idx == len(sheet_names) - 1 and
                      april_idx == len(sheet_names) - 2 and
                      march_idx == len(sheet_names) - 3)
            if correct_order and at_end:
                print(f"PASS: Component 4 — Sheet order correct: {sheet_names} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 — Sheet order incorrect. Got: {sheet_names}. "
                      f"Expected March/April/May as last 3 sheets after template.")
        else:
            print(f"FAIL: Component 4 — Not all copy sheets present, cannot verify order")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Original 'Monthly Template' still intact (0.10 points)
    # Check that template has expected structure (headers, formulas, dimensions)
    # This component only awards points when the copies also exist (i.e., task was attempted)
    # to avoid scoring on initial_env where only template exists
    try:
        copies_exist = all(name in sheet_names for name in ['March', 'April', 'May'])
        if copies_exist:
            # Verify template has expected properties
            has_title = (template['A1'].value == 'Monthly Expense Report')
            has_headers = (template['A4'].value == 'Category' and
                           template['B4'].value == 'Budget')
            has_formulas = (isinstance(template['D5'].value, str) and
                            '=' in str(template['D5'].value))
            if has_title and has_headers and has_formulas:
                print(f"PASS: Component 5 — Original 'Monthly Template' intact with expected content (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 5 — Template content seems modified. "
                      f"title={template['A1'].value}, headers={template['A4'].value}, "
                      f"formula={template['D5'].value}")
        else:
            print(f"FAIL: Component 5 — Copies don't exist, template integrity check skipped "
                  f"(only meaningful after task completion)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
