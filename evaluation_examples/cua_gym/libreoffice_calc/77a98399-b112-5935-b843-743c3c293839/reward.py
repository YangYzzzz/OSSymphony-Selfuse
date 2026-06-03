"""
Reward Script: Track student cohort retention — formulas and stacked bar chart
Task ID: calc_edu_cohort_tracking_044
Domain: libreoffice_calc
Scoring:
  Component 1: Year2 Rate formulas in F2:F7 (=Cn/Bn pattern)          — 0.25 pts
  Component 2: Year3 Rate formulas in G2:G7 (=Dn/Bn pattern)          — 0.15 pts
  Component 3: Year4 Rate formulas in H2:H7 (=En/Bn pattern)          — 0.15 pts
  Component 4: Percentage number format on F, G, H columns             — 0.10 pts
  Component 5: Stacked bar chart present using A1:E7 data              — 0.20 pts
  Component 6: Chart title is 'Cohort Retention Analysis'              — 0.15 pts
  Total: 1.00
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_cohort_tracking_044'


def get_chart_title_text(chart):
    """Extract chart title text from openpyxl chart object."""
    try:
        if chart.title is None:
            return None
        title = chart.title
        # Try rich text structure
        if hasattr(title, 'tx') and title.tx is not None:
            tx = title.tx
            if hasattr(tx, 'rich') and tx.rich is not None:
                for p in tx.rich.p:
                    for r in p.r:
                        if r.t:
                            return r.t
            if hasattr(tx, 'strRef') and tx.strRef is not None:
                return str(tx.strRef)
        # Try direct string
        if isinstance(title, str):
            return title
        return str(title)
    except Exception:
        return None


def count_matching_formulas(ws, col_idx, ref_col_letter, row_range):
    """Count how many cells in col_idx match formula =<ref_col_letter><row>/B<row>."""
    count = 0
    for row in row_range:
        val = ws.cell(row=row, column=col_idx).value
        if val is not None and isinstance(val, str):
            pattern = rf'^\s*=\s*{ref_col_letter}{row}\s*/\s*B{row}\s*$'
            if re.match(pattern, val, re.IGNORECASE):
                count += 1
    return count


def count_pct_formatted_cols(ws, col_indices, row_range):
    """Count how many columns have at least one percentage-formatted cell."""
    cols_with_pct = 0
    for col in col_indices:
        for row in row_range:
            fmt = ws.cell(row=row, column=col).number_format
            if fmt and '%' in fmt:
                cols_with_pct += 1
                break
    return cols_with_pct


def find_bar_chart(ws):
    """Return the first BarChart in the sheet, or None."""
    for chart in ws._charts:
        if type(chart).__name__ == 'BarChart':
            return chart
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the CohortData sheet exists (precondition gate)
    if 'CohortData' not in wb.sheetnames:
        print("CRITICAL: 'CohortData' sheet not found — cannot verify task")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CohortData']
    row_range = range(2, 8)  # rows 2 to 7 (6 cohort rows)

    # Component 1: Year2 Rate formulas in F2:F7 — formula =Cn/Bn (0.25 points)
    # Initial file has None in these cells; golden file has =C2/B2, =C3/B3, ...
    try:
        f2_count = count_matching_formulas(ws, col_idx=6, ref_col_letter='C', row_range=row_range)
        if f2_count == 6:
            print(f"PASS: Component 1 — All 6 Year2 Rate formulas in F2:F7 correct (=Cn/Bn) (0.25 pts)")
            total_score += 0.25
        elif f2_count > 0:
            partial = round(0.25 * f2_count / 6, 4)
            print(f"PARTIAL: Component 1 — {f2_count}/6 Year2 Rate formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No Year2 Rate formulas found in F2:F7")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Year3 Rate formulas in G2:G7 — formula =Dn/Bn (0.15 points)
    try:
        g2_count = count_matching_formulas(ws, col_idx=7, ref_col_letter='D', row_range=row_range)
        if g2_count == 6:
            print(f"PASS: Component 2 — All 6 Year3 Rate formulas in G2:G7 correct (=Dn/Bn) (0.15 pts)")
            total_score += 0.15
        elif g2_count > 0:
            partial = round(0.15 * g2_count / 6, 4)
            print(f"PARTIAL: Component 2 — {g2_count}/6 Year3 Rate formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No Year3 Rate formulas found in G2:G7")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Year4 Rate formulas in H2:H7 — formula =En/Bn (0.15 points)
    try:
        h2_count = count_matching_formulas(ws, col_idx=8, ref_col_letter='E', row_range=row_range)
        if h2_count == 6:
            print(f"PASS: Component 3 — All 6 Year4 Rate formulas in H2:H7 correct (=En/Bn) (0.15 pts)")
            total_score += 0.15
        elif h2_count > 0:
            partial = round(0.15 * h2_count / 6, 4)
            print(f"PARTIAL: Component 3 — {h2_count}/6 Year4 Rate formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No Year4 Rate formulas found in H2:H7")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Percentage number format for columns F, G, H (0.10 points)
    # Initial file has 'General' format; golden file has '0.00%' format
    try:
        cols_with_pct = count_pct_formatted_cols(ws, col_indices=[6, 7, 8], row_range=row_range)
        if cols_with_pct == 3:
            print(f"PASS: Component 4 — All 3 rate columns (F, G, H) have percentage format (0.10 pts)")
            total_score += 0.10
        elif cols_with_pct > 0:
            partial = round(0.10 * cols_with_pct / 3, 4)
            print(f"PARTIAL: Component 4 — {cols_with_pct}/3 rate columns have percentage format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No percentage format found in columns F, G, H")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Stacked bar chart present using data from A1:E7 (0.20 points)
    # Initial file has no chart; golden file has 1 stacked BarChart with 4 series
    try:
        bar_chart = find_bar_chart(ws)
        if bar_chart is None:
            print(f"FAIL: Component 5 — No BarChart found in CohortData sheet")
        elif len(bar_chart.series) == 4:
            print(f"PASS: Component 5 — Stacked bar chart with 4 series present (0.20 pts)")
            total_score += 0.20
        elif len(bar_chart.series) > 0:
            print(f"PARTIAL: Component 5 — BarChart exists but has {len(bar_chart.series)} series (expected 4); awarding partial (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — BarChart found but has no series")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Chart title is 'Cohort Retention Analysis' (0.15 points)
    try:
        title_text = None
        for chart in ws._charts:
            title_text = get_chart_title_text(chart)
            if title_text is not None:
                break
        title_matches = (
            title_text is not None
            and 'cohort retention analysis' in title_text.lower()
        )
        if title_matches:
            print(f"PASS: Component 6 — Chart title is '{title_text}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 6 — Chart title 'Cohort Retention Analysis' not found; found: {repr(title_text)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
