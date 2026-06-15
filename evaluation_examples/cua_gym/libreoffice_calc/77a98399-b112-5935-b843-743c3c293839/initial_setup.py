"""
Initial Setup: Track student cohort retention from freshman to senior year.
Task ID: calc_edu_cohort_tracking_044
Domain: libreoffice_calc

Creates CohortData sheet with 6 cohorts (2019-2024), headcounts in B-E,
and empty retention rate columns F, G, H (to be filled by the agent).
No chart exists yet (the agent must create it).
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_cohort_tracking_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: CohortData ---
    ws = wb.active
    ws.title = 'CohortData'

    # Headers (Row 1)
    headers = [
        'Cohort Year',
        'Year1 Enrolled',
        'Year2 Retained',
        'Year3 Retained',
        'Year4 Graduated',
        'Year2 Rate',
        'Year3 Rate',
        'Year4 Rate',
    ]
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, hdr in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 28

    # Column widths
    col_widths = [14, 16, 16, 16, 18, 12, 12, 12]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows (Rows 2-7): cohorts 2019-2024
    # Columns A-E only; F, G, H intentionally empty (to be calculated by agent)
    cohort_data = [
        # (Cohort Year, Year1 Enrolled, Year2 Retained, Year3 Retained, Year4 Graduated)
        (2019, 1240, 1112, 1024,  968),
        (2020, 1185, 1043,  952,  891),
        (2021, 1320, 1188, 1096, 1034),
        (2022, 1275, 1148, 1054,  None),   # Year4 not yet available
        (2023, 1390, 1245,  None, None),   # Year3 and Year4 not yet available
        (2024, 1410, None,  None, None),   # Only Year1 available
    ]

    data_align = Alignment(horizontal='center', vertical='center')
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row_idx, (cohort_year, y1, y2, y3, y4) in enumerate(cohort_data, 2):
        values = [cohort_year, y1, y2, y3, y4]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = data_align
            cell.border = data_border

        # F, G, H columns left empty intentionally
        for col_idx in range(6, 9):
            cell = ws.cell(row=row_idx, column=col_idx, value=None)
            cell.alignment = data_align
            cell.border = data_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Contents:')
    print('  Sheet: CohortData')
    print('  Rows 1-7: headers + 6 cohort rows (2019-2024)')
    print('  Columns B-E: headcounts filled')
    print('  Columns F-H: EMPTY (Year2 Rate, Year3 Rate, Year4 Rate)')
    print('  No chart (agent must create)')


create_initial()
