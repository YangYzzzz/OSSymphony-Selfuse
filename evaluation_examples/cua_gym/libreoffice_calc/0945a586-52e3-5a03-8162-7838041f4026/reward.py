"""
Reward Script: Property maintenance schedule and work order tracker
Task ID: calc_grs_086
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Conditional formatting on Preventive Maintenance (3 rules: overdue red, due-within-7-days orange, upcoming green)
  Component 2 (0.30): Monthly Cost Summary sheet exists with correct structure and SUM formulas
  Component 3 (0.25): Charts sheet exists with priority/status data table and at least 1 chart
  Component 4 (0.15): Chart is a bar chart titled "Work Orders by Priority and Status" with 3 series
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_086'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # =========================================================================
    # Component 1: Conditional formatting on Preventive Maintenance (0.30 pts)
    # The task requires: overdue in red, due within 7 days in orange, upcoming in green.
    # Initial env has NO conditional formatting; golden adds 3 rules on column D.
    # =========================================================================
    try:
        if 'Preventive Maintenance' not in sheet_names:
            print("FAIL: Component 1 — 'Preventive Maintenance' sheet not found")
        else:
            ws_pm = wb['Preventive Maintenance']
            cf_rules = []
            for cf in ws_pm.conditional_formatting:
                for rule in cf.rules:
                    cf_rules.append(rule)

            if len(cf_rules) >= 3:
                # Check for the three expected conditional formatting colors
                cf_colors = set()
                for rule in cf_rules:
                    if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                        try:
                            rgb = rule.dxf.fill.fgColor.rgb
                            if rgb:
                                cf_colors.add(rgb.upper())
                        except Exception:
                            pass

                # Expected: FFFF0000 (red), FFFF8C00 (orange), FF00B050 (green)
                has_red = any('FF0000' in c for c in cf_colors)
                has_orange = any(c in ('FFFF8C00', 'FFFFA500', 'FFFF6600', 'FFFFC000') or 'FF8C' in c for c in cf_colors)
                has_green = any('00B050' in c or '00FF00' in c or '008000' in c for c in cf_colors)

                matches = sum([has_red, has_orange, has_green])
                if matches >= 3:
                    print(f"PASS: Component 1 — 3 conditional formatting rules with red/orange/green colors (0.30 pts)")
                    total_score += 0.30
                elif matches >= 2:
                    print(f"PARTIAL: Component 1 — {matches}/3 expected colors found in CF rules (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 1 — CF rules found but colors don't match. Colors: {cf_colors}")
            elif len(cf_rules) >= 1:
                print(f"PARTIAL: Component 1 — only {len(cf_rules)} CF rules found (need 3) (0.10 pts)")
                total_score += 0.10
            else:
                print("FAIL: Component 1 — no conditional formatting rules found on Preventive Maintenance")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Monthly Cost Summary sheet (0.30 pts)
    # Must exist with headers and SUM formulas in the Total row.
    # Initial env does NOT have this sheet.
    # =========================================================================
    try:
        # Find a sheet matching "Monthly Cost Summary" or similar
        cost_sheet = None
        for sn in sheet_names:
            if 'cost' in sn.lower() and 'summar' in sn.lower():
                cost_sheet = sn
                break
            elif 'monthly' in sn.lower() and 'cost' in sn.lower():
                cost_sheet = sn
                break
            elif 'monthly' in sn.lower() and 'summar' in sn.lower():
                cost_sheet = sn
                break

        if cost_sheet is None:
            print("FAIL: Component 2 — No 'Monthly Cost Summary' sheet found")
        else:
            ws_mcs = wb[cost_sheet]
            comp2_score = 0.0

            # Sub-check 2a: Sheet has appropriate headers (0.10)
            headers = [ws_mcs.cell(row=1, column=c).value for c in range(1, ws_mcs.max_column + 1)]
            header_strs = [str(h).lower() if h else '' for h in headers]
            has_month_header = any('month' in h for h in header_strs)
            has_cost_header = any('cost' in h for h in header_strs)
            if has_month_header and has_cost_header and ws_mcs.max_row >= 3:
                comp2_score += 0.10
                print(f"PASS: Component 2a — Cost summary sheet has valid headers and data rows (0.10 pts)")
            else:
                print(f"FAIL: Component 2a — headers={headers}, max_row={ws_mcs.max_row}")

            # Sub-check 2b: Has at least one SUM formula (0.10)
            has_sum = False
            for r in range(1, ws_mcs.max_row + 1):
                for c in range(1, ws_mcs.max_column + 1):
                    val = ws_mcs.cell(row=r, column=c).value
                    if isinstance(val, str) and '=SUM' in val.upper():
                        has_sum = True
                        break
                if has_sum:
                    break
            if has_sum:
                comp2_score += 0.10
                print(f"PASS: Component 2b — SUM formula found in cost summary (0.10 pts)")
            else:
                print("FAIL: Component 2b — no SUM formula found in cost summary sheet")

            # Sub-check 2c: Has numeric cost data (not just headers) (0.10)
            has_numeric_cost = False
            for r in range(2, ws_mcs.max_row + 1):
                for c in range(2, ws_mcs.max_column + 1):
                    val = ws_mcs.cell(row=r, column=c).value
                    if isinstance(val, (int, float)) and val > 0:
                        has_numeric_cost = True
                        break
                if has_numeric_cost:
                    break
            if has_numeric_cost:
                comp2_score += 0.10
                print(f"PASS: Component 2c — numeric cost data present (0.10 pts)")
            else:
                print("FAIL: Component 2c — no numeric cost data in summary sheet")

            total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Charts sheet with priority/status data table (0.25 pts)
    # Must have a sheet with priority breakdown data and at least one chart.
    # Initial env does NOT have a Charts sheet or any charts.
    # =========================================================================
    try:
        # Find a sheet with chart(s)
        chart_sheet = None
        for sn in sheet_names:
            ws_check = wb[sn]
            if len(ws_check._charts) > 0:
                chart_sheet = sn
                break

        if chart_sheet is None:
            # Also check for a sheet named Charts even without embedded chart objects
            for sn in sheet_names:
                if 'chart' in sn.lower():
                    chart_sheet = sn
                    break

        if chart_sheet is None:
            print("FAIL: Component 3 — No sheet with charts found")
        else:
            ws_ch = wb[chart_sheet]
            comp3_score = 0.0

            # Sub-check 3a: Has at least one chart object (0.15)
            if len(ws_ch._charts) >= 1:
                comp3_score += 0.15
                print(f"PASS: Component 3a — {len(ws_ch._charts)} chart(s) found on '{chart_sheet}' (0.15 pts)")
            else:
                print(f"FAIL: Component 3a — no chart objects found on '{chart_sheet}'")

            # Sub-check 3b: Has priority/status data table (0.10)
            # Look for priority values in the sheet data
            priority_found = False
            for r in range(1, ws_ch.max_row + 1):
                for c in range(1, ws_ch.max_column + 1):
                    val = ws_ch.cell(row=r, column=c).value
                    if val and str(val).lower() in ('emergency', 'high', 'medium', 'low', 'priority'):
                        priority_found = True
                        break
                if priority_found:
                    break

            if priority_found:
                comp3_score += 0.10
                print(f"PASS: Component 3b — priority/status data table present (0.10 pts)")
            else:
                print("FAIL: Component 3b — no priority/status data found on chart sheet")

            total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Chart properties (0.15 pts)
    # The chart should be a bar/column chart about work orders by priority/status.
    # =========================================================================
    try:
        # Find any chart across all sheets
        chart_obj = None
        for sn in sheet_names:
            ws_check = wb[sn]
            if len(ws_check._charts) > 0:
                chart_obj = ws_check._charts[0]
                break

        if chart_obj is None:
            print("FAIL: Component 4 — No chart object to inspect")
        else:
            comp4_score = 0.0
            chart_class = type(chart_obj).__name__

            # Sub-check 4a: Chart is a bar/column type (0.07)
            if chart_class in ('BarChart', 'BarChart3D'):
                comp4_score += 0.07
                print(f"PASS: Component 4a — chart is BarChart type (0.07 pts)")
            else:
                print(f"FAIL: Component 4a — chart type is {chart_class}, expected BarChart")

            # Sub-check 4b: Chart has a title containing relevant keywords (0.08)
            title_text = ''
            if chart_obj.title:
                try:
                    # Title can be a string or a Title object
                    if isinstance(chart_obj.title, str):
                        title_text = chart_obj.title
                    else:
                        # Extract from rich text
                        for p in chart_obj.title.tx.rich.paragraphs:
                            for run in p.r:
                                title_text += run.t
                except Exception:
                    title_text = str(chart_obj.title)

            title_lower = title_text.lower()
            if ('work order' in title_lower or 'priority' in title_lower or 'status' in title_lower):
                comp4_score += 0.08
                print(f"PASS: Component 4b — chart title '{title_text}' contains relevant keywords (0.08 pts)")
            elif title_text:
                comp4_score += 0.04
                print(f"PARTIAL: Component 4b — chart has title '{title_text}' but missing keywords (0.04 pts)")
            else:
                print("FAIL: Component 4b — chart has no title")

            total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
