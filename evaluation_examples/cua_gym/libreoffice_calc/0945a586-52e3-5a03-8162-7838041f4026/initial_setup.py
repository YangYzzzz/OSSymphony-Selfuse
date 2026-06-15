"""
Initial Setup: Property maintenance schedule and work order tracker
Task ID: calc_grs_086
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========================================================
    # Sheet 1: Preventive Maintenance Schedule
    # ========================================================
    ws1 = wb.active
    ws1.title = "Preventive Maintenance"

    headers1 = [
        "Equipment/Area", "Maintenance Task", "Frequency",
        "Next Due Date", "Last Completed", "Responsible Party",
        "Estimated Hours", "Cost"
    ]

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Realistic maintenance data for an 8-floor commercial office building
    today = date(2026, 4, 2)
    maintenance_data = [
        ["HVAC System - Floors 1-4", "Filter replacement and coil cleaning", "Monthly",
         date(2026, 4, 15), date(2026, 3, 15), "Mike Torres", 3.0, 450.00],
        ["HVAC System - Floors 5-8", "Filter replacement and coil cleaning", "Monthly",
         date(2026, 4, 18), date(2026, 3, 18), "Mike Torres", 3.0, 450.00],
        ["Elevator Bank A", "Full inspection and lubrication", "Quarterly",
         date(2026, 3, 28), date(2025, 12, 28), "Otis Elevator Co.", 4.0, 1200.00],
        ["Elevator Bank B", "Full inspection and lubrication", "Quarterly",
         date(2026, 6, 15), date(2026, 3, 15), "Otis Elevator Co.", 4.0, 1200.00],
        ["Fire Suppression System", "Sprinkler head inspection and flow test", "Annual",
         date(2026, 7, 1), date(2025, 7, 1), "SafeGuard Fire Protection", 8.0, 3500.00],
        ["Parking Garage Lighting", "Replace burned-out fixtures and clean lenses", "Weekly",
         date(2026, 4, 7), date(2026, 3, 31), "David Nakamura", 1.5, 85.00],
        ["Lobby and Common Areas", "Floor buffing and deep clean", "Weekly",
         date(2026, 4, 5), date(2026, 3, 29), "CleanPro Services", 2.0, 200.00],
        ["Roof Membrane", "Inspect seams, drains, and flashing", "Quarterly",
         date(2026, 3, 20), date(2025, 12, 20), "Summit Roofing", 3.0, 800.00],
        ["Emergency Generator", "Load bank test and fluid check", "Monthly",
         date(2026, 4, 10), date(2026, 3, 10), "PowerGen Solutions", 2.5, 350.00],
        ["Plumbing - Main Risers", "Inspect valves and check for leaks", "Quarterly",
         date(2026, 5, 1), date(2026, 2, 1), "Reliable Plumbing", 3.0, 600.00],
        ["Window Washing - Exterior", "Full exterior wash all 8 floors", "Quarterly",
         date(2026, 5, 15), date(2026, 2, 15), "ClearView Window Co.", 16.0, 4800.00],
        ["Security System", "Camera and access panel check", "Monthly",
         date(2026, 3, 25), date(2026, 2, 25), "SecureTech Inc.", 2.0, 300.00],
        ["Landscaping", "Trim hedges, mow lawn, seasonal planting", "Weekly",
         date(2026, 4, 6), date(2026, 3, 30), "GreenScape Landscaping", 4.0, 350.00],
        ["Boiler System", "Inspect burners, clean heat exchanger", "Annual",
         date(2026, 10, 1), date(2025, 10, 1), "HeatMaster Services", 6.0, 2200.00],
    ]

    for r, row_data in enumerate(maintenance_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (4, 5):  # date columns
                cell.number_format = 'yyyy-mm-dd'
            elif c == 8:  # cost
                cell.number_format = '$#,##0.00'
            elif c == 7:  # hours
                cell.number_format = '0.0'

    # Frequency dropdown validation
    freq_dv = DataValidation(
        type="list",
        formula1='"Daily,Weekly,Monthly,Quarterly,Annual"',
        allow_blank=True,
        showDropDown=False,
    )
    freq_dv.error = "Please select a valid frequency"
    freq_dv.errorTitle = "Invalid Frequency"
    freq_dv.add(f"C2:C100")
    ws1.add_data_validation(freq_dv)

    # Column widths
    col_widths1 = {"A": 30, "B": 38, "C": 14, "D": 15, "E": 15, "F": 24, "G": 14, "H": 12}
    for col_letter, width in col_widths1.items():
        ws1.column_dimensions[col_letter].width = width

    ws1.freeze_panes = "A2"

    # ========================================================
    # Sheet 2: Work Orders
    # ========================================================
    ws2 = wb.create_sheet("Work Orders")

    headers2 = [
        "WO #", "Request Date", "Requested By", "Description", "Priority",
        "Assigned Technician", "Status", "Start Date", "Completion Date",
        "Labor Hours", "Parts Cost", "Total Cost"
    ]

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    work_orders = [
        ["WO-2026-001", date(2026, 1, 8), "Tenant - Floor 3 (Apex Legal)", "Ceiling tile water stain in conference room 3B",
         "High", "David Nakamura", "Completed", date(2026, 1, 8), date(2026, 1, 9), 2.5, 45.00, 170.00],
        ["WO-2026-002", date(2026, 1, 12), "Building Manager", "Replace worn carpet tiles in lobby entrance",
         "Medium", "CleanPro Services", "Completed", date(2026, 1, 15), date(2026, 1, 16), 6.0, 890.00, 1190.00],
        ["WO-2026-003", date(2026, 1, 20), "Tenant - Floor 6 (DataSync Corp)", "Flickering lights in open office area",
         "Medium", "David Nakamura", "Completed", date(2026, 1, 21), date(2026, 1, 21), 1.5, 120.00, 195.00],
        ["WO-2026-004", date(2026, 1, 25), "Security Team", "Parking garage gate motor malfunction",
         "Emergency", "PowerGen Solutions", "Completed", date(2026, 1, 25), date(2026, 1, 25), 3.0, 650.00, 800.00],
        ["WO-2026-005", date(2026, 2, 3), "Tenant - Floor 2 (Meridian Finance)", "Thermostat not responding in suite 210",
         "High", "Mike Torres", "Completed", date(2026, 2, 3), date(2026, 2, 4), 2.0, 180.00, 280.00],
        ["WO-2026-006", date(2026, 2, 10), "Janitorial Staff", "Restroom Floor 4 - toilet running continuously",
         "Medium", "Reliable Plumbing", "Completed", date(2026, 2, 11), date(2026, 2, 11), 1.0, 35.00, 85.00],
        ["WO-2026-007", date(2026, 2, 18), "Tenant - Floor 7 (Vertex Analytics)", "Window blind stuck in closed position",
         "Low", "David Nakamura", "Completed", date(2026, 2, 20), date(2026, 2, 20), 0.5, 0.00, 25.00],
        ["WO-2026-008", date(2026, 2, 22), "Building Manager", "Emergency exit sign burned out - Floor 5 east stairwell",
         "Emergency", "David Nakamura", "Completed", date(2026, 2, 22), date(2026, 2, 22), 0.5, 45.00, 70.00],
        ["WO-2026-009", date(2026, 3, 1), "Tenant - Floor 1 (Brew & Bean Cafe)", "Grease trap needs cleaning",
         "High", "Reliable Plumbing", "Completed", date(2026, 3, 2), date(2026, 3, 2), 2.0, 0.00, 100.00],
        ["WO-2026-010", date(2026, 3, 5), "Tenant - Floor 4 (Sterling Consulting)", "Door lock keypad not accepting code",
         "High", "SecureTech Inc.", "Completed", date(2026, 3, 5), date(2026, 3, 6), 1.5, 220.00, 295.00],
        ["WO-2026-011", date(2026, 3, 10), "Building Manager", "Parking lot pothole repair needed near entrance",
         "Medium", "Summit Roofing", "In Progress", date(2026, 3, 12), None, 0.0, 0.00, 0.00],
        ["WO-2026-012", date(2026, 3, 15), "Tenant - Floor 8 (Horizon Media)", "AC blowing warm air in server room",
         "Emergency", "Mike Torres", "In Progress", date(2026, 3, 15), None, 1.0, 0.00, 50.00],
        ["WO-2026-013", date(2026, 3, 18), "Janitorial Staff", "Elevator B - unusual grinding noise",
         "High", "Otis Elevator Co.", "Open", None, None, 0.0, 0.00, 0.00],
        ["WO-2026-014", date(2026, 3, 22), "Tenant - Floor 5 (NovaTech Labs)", "Stained ceiling tiles from roof leak",
         "Medium", "Summit Roofing", "Open", None, None, 0.0, 0.00, 0.00],
        ["WO-2026-015", date(2026, 3, 28), "Security Team", "Camera 12 offline - parking level 2",
         "Low", "SecureTech Inc.", "Open", None, None, 0.0, 0.00, 0.00],
        ["WO-2026-016", date(2026, 3, 30), "Tenant - Floor 3 (Apex Legal)", "Squeaky door hinge - main entrance suite 305",
         "Low", "David Nakamura", "Open", None, None, 0.0, 0.00, 0.00],
    ]

    for r, row_data in enumerate(work_orders, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (2, 8, 9):  # date columns
                cell.number_format = 'yyyy-mm-dd'
            elif c in (11, 12):  # cost columns
                cell.number_format = '$#,##0.00'
            elif c == 10:  # hours
                cell.number_format = '0.0'

    # Priority dropdown validation
    priority_dv = DataValidation(
        type="list",
        formula1='"Emergency,High,Medium,Low"',
        allow_blank=True,
        showDropDown=False,
    )
    priority_dv.error = "Please select a valid priority"
    priority_dv.errorTitle = "Invalid Priority"
    priority_dv.add("E2:E100")
    ws2.add_data_validation(priority_dv)

    # Column widths
    col_widths2 = {
        "A": 16, "B": 14, "C": 32, "D": 48, "E": 12,
        "F": 22, "G": 14, "H": 14, "I": 16, "J": 12, "K": 12, "L": 12
    }
    for col_letter, width in col_widths2.items():
        ws2.column_dimensions[col_letter].width = width

    ws2.freeze_panes = "A2"

    # ========================================================
    # NO conditional formatting, NO cost summary, NO charts
    # (Task asks the agent to create these)
    # ========================================================

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
