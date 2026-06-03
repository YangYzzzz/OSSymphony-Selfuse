"""
Initial Setup: Balance sheet raw financial data for Meridian Holdings LLC
Task ID: calc_gpm_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'BalanceSheet'

    # --- Company Header (merged, styled) ---
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Meridian Holdings LLC'
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:D2')
    ws['A2'] = 'Balance Sheet as of March 31, 2026'
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(italic=True)

    # --- ASSETS section ---
    ws['A4'] = 'ASSETS'
    ws['A4'].font = Font(bold=True, underline='single')

    ws['A5'] = 'Current Assets'
    ws['A5'].font = Font(bold=True)

    ws['A6'] = '  Cash'
    ws['C6'] = 125000

    ws['A7'] = '  Accounts Receivable'
    ws['C7'] = 89000

    ws['A8'] = '  Inventory'
    ws['C8'] = 67000

    ws['A9'] = 'Total Current Assets'
    ws['A9'].font = Font(bold=True)
    # D9 intentionally left empty (agent must add =SUM(C6:C8))

    ws['A10'] = 'Non-Current Assets'
    ws['A10'].font = Font(bold=True)

    ws['A11'] = '  Property & Equipment'
    ws['C11'] = 350000

    ws['A12'] = '  Less: Accumulated Depreciation'
    ws['C12'] = -85000

    ws['A13'] = 'Total Non-Current'
    # D13 intentionally left empty (agent must add =SUM(C11:C12))

    ws['A14'] = 'TOTAL ASSETS'
    ws['A14'].font = Font(bold=True)
    # D14 intentionally left empty (agent must add =D9+D13)

    # --- LIABILITIES section ---
    ws['A16'] = 'LIABILITIES'
    ws['A16'].font = Font(bold=True, underline='single')

    ws['A17'] = '  Accounts Payable'
    ws['C17'] = 45000

    ws['A18'] = '  Notes Payable'
    ws['C18'] = 150000

    ws['A19'] = 'Total Liabilities'
    ws['A19'].font = Font(bold=True)
    # D19 intentionally left empty (agent must add =SUM(C17:C18))

    # --- EQUITY section ---
    ws['A20'] = 'EQUITY'
    ws['A20'].font = Font(bold=True, underline='single')

    ws['A21'] = '  Common Stock'
    ws['C21'] = 200000

    ws['A22'] = '  Retained Earnings'
    ws['C22'] = 151000

    ws['A23'] = 'Total Equity'
    # D23 intentionally left empty (agent must add =SUM(C21:C22))

    ws['A24'] = 'TOTAL LIABILITIES & EQUITY'
    ws['A24'].font = Font(bold=True)
    # D24 intentionally left empty (agent must add =D19+D23)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 5
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # NOTE: No formulas in column D, no top borders on subtotals,
    # no double underlines on totals, no dollar number formatting.
    # The agent must add all of these.

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
