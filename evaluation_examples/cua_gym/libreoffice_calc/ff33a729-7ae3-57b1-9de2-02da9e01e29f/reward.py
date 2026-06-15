"""
Reward Script: Balance sheet with assets, liabilities, and equity sections
Task ID: calc_gpm_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): D-column formulas for subtotals and totals
  Component 2 (0.30): Dollar number format ($#,##0) on monetary cells
  Component 3 (0.20): Top borders on subtotal cells (D9, D13, D19, D23)
  Component 4 (0.20): Double bottom borders on total cells (D14, D24)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_008'


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'BalanceSheet' must exist
    if 'BalanceSheet' not in wb.sheetnames:
        print(f"FAIL: Sheet 'BalanceSheet' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['BalanceSheet']

    # ----------------------------------------------------------------
    # Component 1: D-column formulas for subtotals and totals (0.30 pts)
    # These formulas do NOT exist in the initial file (D column is empty).
    # ----------------------------------------------------------------
    try:
        expected_formulas = {
            'D9':  '=SUM(C6:C8)',
            'D13': '=SUM(C11:C12)',
            'D14': '=D9+D13',
            'D19': '=SUM(C17:C18)',
            'D23': '=SUM(C21:C22)',
            'D24': '=D19+D23',
        }
        formula_pass_count = 0
        for coord, expected in expected_formulas.items():
            val = ws[coord].value
            if val is not None and isinstance(val, str):
                # Normalize: remove spaces, uppercase
                actual_norm = val.upper().replace(" ", "")
                expected_norm = expected.upper().replace(" ", "")
                if actual_norm == expected_norm:
                    formula_pass_count += 1
                    print(f"  PASS: {coord} has formula {val}")
                else:
                    print(f"  FAIL: {coord} expected formula {expected}, found {val}")
            else:
                print(f"  FAIL: {coord} expected formula {expected}, found {repr(val)}")

        if formula_pass_count == len(expected_formulas):
            print(f"PASS: Component 1 -- All 6 D-column formulas correct (0.30 pts)")
            total_score += 0.30
        elif formula_pass_count >= 4:
            partial = round(0.30 * formula_pass_count / len(expected_formulas), 2)
            print(f"PARTIAL: Component 1 -- {formula_pass_count}/6 formulas correct ({partial} pts)")
            total_score += partial
        elif formula_pass_count >= 1:
            partial = round(0.30 * formula_pass_count / len(expected_formulas), 2)
            print(f"PARTIAL: Component 1 -- {formula_pass_count}/6 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No D-column formulas found (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # ----------------------------------------------------------------
    # Component 2: Dollar number format $#,##0 on monetary cells (0.30 pts)
    # Initial file has 'General' format on all C-column numbers.
    # Golden has '$#,##0' on C6:C8, C11:C12, C17:C18, C21:C22 and
    # D9, D13, D14, D19, D23, D24.
    # ----------------------------------------------------------------
    try:
        dollar_cells = [
            'C6', 'C7', 'C8', 'C11', 'C12', 'C17', 'C18', 'C21', 'C22',
            'D9', 'D13', 'D14', 'D19', 'D23', 'D24'
        ]
        dollar_pass_count = 0
        for coord in dollar_cells:
            nfmt = ws[coord].number_format
            # Accept any format containing '$' and '#' as dollar format
            if nfmt and '$' in str(nfmt):
                dollar_pass_count += 1
            else:
                print(f"  FAIL: {coord} number_format is {repr(nfmt)}, expected dollar format")

        if dollar_pass_count == len(dollar_cells):
            print(f"PASS: Component 2 -- All {len(dollar_cells)} cells have dollar format (0.30 pts)")
            total_score += 0.30
        elif dollar_pass_count >= 1:
            partial = round(0.30 * dollar_pass_count / len(dollar_cells), 2)
            print(f"PARTIAL: Component 2 -- {dollar_pass_count}/{len(dollar_cells)} cells have dollar format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No dollar formatted cells found (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # ----------------------------------------------------------------
    # Component 3: Top borders on subtotal D cells (0.20 pts)
    # D9, D13, D19, D23 should have thin top border (absent in initial).
    # ----------------------------------------------------------------
    try:
        top_border_cells = ['D9', 'D13', 'D19', 'D23']
        top_border_pass = 0
        for coord in top_border_cells:
            cell = ws[coord]
            top_style = None
            try:
                if cell.border and cell.border.top:
                    top_style = cell.border.top.style
            except Exception:
                pass
            if top_style is not None and top_style != 'none':
                top_border_pass += 1
                print(f"  PASS: {coord} has top border style={top_style}")
            else:
                print(f"  FAIL: {coord} has no top border (style={top_style})")

        if top_border_pass == len(top_border_cells):
            print(f"PASS: Component 3 -- All 4 subtotal cells have top borders (0.20 pts)")
            total_score += 0.20
        elif top_border_pass >= 1:
            partial = round(0.20 * top_border_pass / len(top_border_cells), 2)
            print(f"PARTIAL: Component 3 -- {top_border_pass}/4 subtotal cells have top borders ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No subtotal cells have top borders (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # ----------------------------------------------------------------
    # Component 4: Double bottom borders on total D cells (0.20 pts)
    # D14, D24 should have double bottom border (absent in initial).
    # ----------------------------------------------------------------
    try:
        dbl_border_cells = ['D14', 'D24']
        dbl_border_pass = 0
        for coord in dbl_border_cells:
            cell = ws[coord]
            bot_style = None
            try:
                if cell.border and cell.border.bottom:
                    bot_style = cell.border.bottom.style
            except Exception:
                pass
            if bot_style == 'double':
                dbl_border_pass += 1
                print(f"  PASS: {coord} has double bottom border")
            else:
                print(f"  FAIL: {coord} bottom border style={bot_style}, expected 'double'")

        if dbl_border_pass == len(dbl_border_cells):
            print(f"PASS: Component 4 -- Both total cells have double bottom borders (0.20 pts)")
            total_score += 0.20
        elif dbl_border_pass >= 1:
            partial = round(0.20 * dbl_border_pass / len(dbl_border_cells), 2)
            print(f"PARTIAL: Component 4 -- {dbl_border_pass}/2 total cells have double bottom borders ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 -- No total cells have double bottom borders (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
