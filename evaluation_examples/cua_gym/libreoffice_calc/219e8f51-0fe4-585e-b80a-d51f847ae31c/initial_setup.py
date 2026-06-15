"""
Initial Setup: Apply conditional formatting rules to sales data
Task ID: calc_ggf_049
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # --- Headers ---
    headers = ["Rep ID", "Name", "Region", "Target", "Actual Sales"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16

    # --- Data: 100 sales reps (rows 2-101) ---
    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei", "Carlos",
        "Olivia", "Kwame", "Aisha", "Thomas", "Yuki", "Robert", "Fatima",
        "Michael", "Lena", "Andre", "Sofia", "Daniel", "Amara", "Patrick",
        "Nina", "Raj", "Emily", "Hassan", "Julia", "Wei", "Catherine", "Omar",
        "Hannah", "Victor", "Ingrid", "Samuel", "Zara", "Brian", "Keiko",
        "Antonio", "Rachel", "Dmitri", "Grace", "Felix", "Nadia", "Lukas",
        "Bianca", "Tariq", "Megan", "Hiroshi", "Claudia", "Sean",
    ]
    last_names = [
        "Chen", "Johnson", "Rivera", "O'Brien", "Patel", "Kim", "Zhang",
        "Lopez", "Williams", "Osei", "Mohammed", "Anderson", "Tanaka",
        "Taylor", "Hassan", "Brown", "Mueller", "Dupont", "Garcia", "Lee",
        "Okafor", "Murphy", "Petrov", "Sharma", "Davis", "Ali", "Fischer",
        "Wang", "Martin", "Abdel", "Baker", "Santos", "Johansson", "Nguyen",
        "Khan", "Clarke", "Yamamoto", "Rossi", "Cohen", "Volkov", "Achebe",
        "Costa", "Svensson", "Park", "Romano", "Singh", "Campbell", "Ito",
        "Laurent", "Walsh",
    ]
    regions = ["Northeast", "Southeast", "Midwest", "West", "Central",
               "Pacific", "Mountain", "Atlantic"]

    random.seed(42)  # reproducible data

    data_font = Font(name="Calibri", size=11)
    num_align = Alignment(horizontal="right")

    for i in range(100):
        row = i + 2
        rep_id = f"SR-{1001 + i}"
        name = f"{first_names[i % len(first_names)]} {last_names[i % len(last_names)]}"
        region = regions[i % len(regions)]
        target = round(random.uniform(40000, 120000), 2)
        # Actual sales: varied range so there are clear top and bottom performers
        actual = round(random.uniform(18000, 155000), 2)

        ws.cell(row=row, column=1, value=rep_id).font = data_font
        ws.cell(row=row, column=2, value=name).font = data_font
        ws.cell(row=row, column=3, value=region).font = data_font

        target_cell = ws.cell(row=row, column=4, value=target)
        target_cell.font = data_font
        target_cell.number_format = '$#,##0.00'
        target_cell.alignment = num_align

        sales_cell = ws.cell(row=row, column=5, value=actual)
        sales_cell.font = data_font
        sales_cell.number_format = '$#,##0.00'
        sales_cell.alignment = num_align

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting - that's the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
