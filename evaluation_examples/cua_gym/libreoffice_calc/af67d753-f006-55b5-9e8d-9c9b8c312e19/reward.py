"""
Reward Script: Apply a double-line bottom border beneath the totals row (row 12) in columns A through E
Task ID: calc_fmt_border_double_015
Domain: libreoffice_calc

Scoring:
  Component 1: A12 has double bottom border         — 0.2 pts
  Component 2: B12 has double bottom border         — 0.2 pts
  Component 3: C12 has double bottom border         — 0.2 pts
  Component 4: D12 has double bottom border         — 0.2 pts
  Component 5: E12 has double bottom border         — 0.2 pts
  Total: 1.0

NOTE: The initial file has NO borders on any cells. Only the golden file has
double bottom borders on A12:E12. Each component scores ONLY if the double
bottom border is present — meaning 0.0 on initial, 1.0 on golden.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_border_double_015'
SHEET_NAME = 'Income Statement'
TARGET_ROW = 12
TARGET_COLS = ['A', 'B', 'C', 'D', 'E']
EXPECTED_BORDER_STYLE = 'double'


def has_double_bottom_border(cell):
    """Return True if cell has a double-line bottom border."""
    try:
        b = cell.border
        if b is None:
            return False
        bottom = b.bottom
        if bottom is None:
            return False
        return bottom.border_style == EXPECTED_BORDER_STYLE
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Checks that each of the 5 cells A12:E12 has a double-line bottom border.
    Each cell contributes 0.2 points. Total = 1.0 if all 5 borders are present.
    """
    total_score = 0.0

    # Precondition: Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Components 1-5: Each of A12, B12, C12, D12, E12 must have a double bottom border
    # (0.2 points each)
    for col_letter in TARGET_COLS:
        coord = f"{col_letter}{TARGET_ROW}"
        try:
            cell = ws[coord]
            if has_double_bottom_border(cell):
                actual_style = cell.border.bottom.border_style
                print(f"PASS: {coord} has double bottom border (style='{actual_style}') (+0.2 pts)")
                total_score += 0.2
            else:
                actual_style = None
                try:
                    if cell.border and cell.border.bottom:
                        actual_style = cell.border.bottom.border_style
                except Exception:
                    pass
                print(f"FAIL: {coord} missing double bottom border — found: '{actual_style}'")
        except Exception as e:
            print(f"ERROR: Cannot check border on {coord}: {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.1f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
