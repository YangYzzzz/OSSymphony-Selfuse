"""
Initial Setup: Apply a double-line bottom border beneath the totals row (row 12) in columns A through E
Task ID: calc_fmt_border_double_015
Domain: libreoffice_calc

Creates an Income Statement spreadsheet with no borders. Row 12 is the totals row (Net Income).
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_border_double_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Income Statement'

    # --- Headers (Row 1) ---
    headers = ['Line Item', 'FY2023', 'FY2024', 'FY2025', 'CAGR']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Data rows (Rows 2-11): realistic income statement line items ---
    data = [
        ['Revenue',             1850000, 2120000, 2390000, '13.6%'],
        ['Cost of Goods Sold',   740000,  848000,  956000, '13.6%'],
        ['Gross Profit',        1110000, 1272000, 1434000, '13.6%'],
        ['Research & Development', 185000, 212000, 239000, '13.6%'],
        ['Sales & Marketing',    277500,  318000,  358500, '13.7%'],
        ['General & Administrative', 148000, 169600, 191200, '13.6%'],
        ['Total Operating Expenses', 610500, 699600, 788700, '13.6%'],
        ['Operating Income',    499500,  572400,  645300, '13.5%'],
        ['Interest Expense',     48000,   52000,   58000, '9.9%'],
        ['Income Before Tax',   451500,  520400,  587300, '14.1%'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (2, 3, 4) and isinstance(val, int):
                cell.number_format = '#,##0'

    # --- Row 12: Net Income (totals row) — NO borders per task spec ---
    totals = ['Net Income', 234000, 289000, 341000, '20.8%']
    for c, val in enumerate(totals, 1):
        cell = ws.cell(row=12, column=c, value=val)
        cell.font = Font(bold=True)
        if c in (2, 3, 4):
            cell.number_format = '#,##0'

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10

    # --- Row 1 height ---
    ws.row_dimensions[1].height = 18

    # NOTE: No borders set on any cells — this is the pre-task state.

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
