"""
Initial Setup: Meeting Notes spreadsheet with wrap=False and vertical alignment=BOTTOM for column G
Task ID: calc_fmt_wrap_and_align_notes_079
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_wrap_and_align_notes_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Meeting Notes'

    # --- Headers (Row 1) ---
    headers = ['Date', 'Meeting', 'Attendees', 'Duration', 'Action Items', 'Owner', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 40

    # --- Row heights for data rows ---
    for r in range(2, 26):
        ws.row_dimensions[r].height = 40

    # --- Meeting data (rows 2-25) ---
    data = [
        ('2025-01-06', 'Q1 Planning Kickoff', 'Sarah Chen, Marcus Johnson, Lisa Park', '90 min',
         'Finalize Q1 OKRs by Jan 10; assign budget owners', 'Sarah Chen',
         'Team agreed on three strategic pillars for Q1. Budget discussions deferred to next week. Everyone needs to submit their departmental goals by end of Thursday.'),
        ('2025-01-09', 'Product Roadmap Review', 'James Wu, Priya Sharma, David Lee', '60 min',
         'Update roadmap deck with new features; share with stakeholders by Friday', 'James Wu',
         'Feature prioritization complete. Mobile app enhancements moved to Q2 based on resource availability. Backend API work stays on Q1 track. James to circulate updated roadmap.'),
        ('2025-01-13', 'Weekly Engineering Standup', 'David Lee, Mei Lin, Carlos Rivera', '30 min',
         'Resolve CI/CD pipeline issue before Thursday deployment; update docs', 'David Lee',
         'Pipeline failure traced to misconfigured environment variable. Carlos will push fix today. Deployment window confirmed for Thursday 11pm UTC. Documentation update assigned to Mei.'),
        ('2025-01-15', 'HR Policy Update Session', 'Natalie Brooks, Tom Fischer, Karen Yuen', '45 min',
         'Distribute updated remote work policy to all staff; collect signatures by Jan 22', 'Natalie Brooks',
         'Revised remote work policy now allows up to 3 days WFH per week. New expense reimbursement limits effective February 1. HR will send DocuSign links to everyone by end of day Friday.'),
        ('2025-01-20', 'Customer Success Weekly', 'Rachel Kim, Omar Hassan, Sofia Martini', '60 min',
         'Follow up with Acme Corp on renewal; schedule QBR for GlobalTech', 'Rachel Kim',
         'Acme Corp renewal at risk due to pricing concerns — Rachel to escalate to VP Sales. GlobalTech QBR tentatively set for Feb 5. NPS survey results show 7.8 average, up from 7.2 last quarter.'),
        ('2025-01-22', 'Finance Monthly Close Review', 'Karen Yuen, Brian O\'Sullivan, Lena Vogel', '90 min',
         'Reconcile outstanding invoices; prepare board deck revenue slide', 'Karen Yuen',
         'December close complete with $2.1M revenue, slightly below $2.3M forecast. Two large invoices from November still unpaid — collections team following up. Board deck due Jan 31. Lena to draft revenue slide by Jan 28.'),
        ('2025-01-27', 'Design Sprint Retrospective', 'Mei Lin, Jasper Wong, Alicia Torres', '45 min',
         'Document sprint learnings in Confluence; schedule next sprint kick-off', 'Mei Lin',
         'Sprint delivered three user-tested prototypes for the onboarding flow. Key insight: users want fewer steps in signup. Next sprint starts Feb 3 with focus on dashboard redesign. Confluence page to be updated by Mei before EOD.'),
        ('2025-01-29', 'Security Audit Debrief', 'Carlos Rivera, Tom Fischer, David Lee', '60 min',
         'Patch critical vulnerabilities by Feb 14; schedule pen-test follow-up', 'Carlos Rivera',
         'External audit identified two high-severity vulnerabilities in authentication module. Patches scoped and assigned. Pen-test follow-up scheduled for March 1. Tom to coordinate with compliance team on remediation timeline.'),
        ('2025-02-03', 'Sales Pipeline Review', 'Marcus Johnson, Rachel Kim, Leo Brennan', '75 min',
         'Update CRM pipeline stages; prepare forecast for Feb board update', 'Marcus Johnson',
         'Pipeline totals $4.8M with 68% probability-weighted. Three deals expected to close this month. Leo to update Salesforce with latest call notes. Forecast presentation for board meeting on Feb 12.'),
        ('2025-02-05', 'Marketing Campaign Planning', 'Priya Sharma, Alicia Torres, Ryan Chow', '60 min',
         'Finalize ad creative by Feb 10; submit budget request for Q2 campaign', 'Priya Sharma',
         'Spring campaign theme approved: "Work Smarter Together". Budget request for $85K submitted for Q2. Creative brief shared with agency. Ryan to review landing page copy and provide feedback by Feb 7.'),
        ('2025-02-10', 'Product Weekly Sync', 'James Wu, Mei Lin, Sofia Martini', '45 min',
         'Confirm feature freeze date; update sprint backlog with customer requests', 'James Wu',
         'Feature freeze confirmed for Feb 21. Customer-requested features triaged — five added to backlog, two deferred to Q3. Beta release date moved to March 15 to allow extra QA time. Backlog refinement session booked for Feb 12.'),
        ('2025-02-12', 'Board Preparation Meeting', 'Sarah Chen, Karen Yuen, Marcus Johnson', '120 min',
         'Finalize board deck by Feb 18; prepare executive talking points', 'Sarah Chen',
         'Board deck structure agreed. Revenue slide, product milestone slide, and hiring update to be included. Karen presenting financials, Marcus presenting go-to-market. Sarah to review full deck by Feb 16. Exec talking points to be drafted by Feb 14.'),
        ('2025-02-14', 'Infrastructure Planning Review', 'David Lee, Carlos Rivera, Lena Vogel', '60 min',
         'Submit cloud cost optimization proposal; evaluate migration to Kubernetes', 'David Lee',
         'Cloud spend grew 23% QoQ largely due to staging environment over-provisioning. Auto-scaling policies to be updated by end of month. Kubernetes migration evaluated — full rollout now planned for Q3 to reduce risk. Cost optimization proposal due Feb 21.'),
        ('2025-02-17', 'Customer Advisory Board Call', 'Sarah Chen, Rachel Kim, James Wu', '90 min',
         'Summarize CAB feedback in Confluence; share product roadmap priorities', 'Rachel Kim',
         'Four enterprise customers participated. Top feedback themes: better reporting, API rate limit increases, and SSO support. James to update roadmap with CAB priorities. Confluence summary to be published within 48 hours. Next CAB call in April.'),
        ('2025-02-19', 'Recruiting Pipeline Update', 'Natalie Brooks, Leo Brennan, Karen Yuen', '45 min',
         'Schedule final interviews for three open roles; extend offers by Feb 28', 'Natalie Brooks',
         'Senior Engineer role has two strong finalists — final interviews Feb 21 and 22. Product Manager role still lacks qualified candidates; JD to be revised. Sales AE role — offer extended to top candidate, response expected by Feb 24. Karen to approve compensation packages.'),
        ('2025-02-24', 'OKR Mid-Quarter Check-in', 'Sarah Chen, All Department Heads', '90 min',
         'Flag off-track OKRs; prepare remediation plans for at-risk KRs', 'Sarah Chen',
         'Engineering 85% on track. Sales at 72% — pipeline concerns noted. Marketing 91% — campaigns performing well. HR 60% on track due to delayed recruiting. Finance OKRs solid. Three KRs flagged as at-risk; owners to submit recovery plans by Feb 28.'),
        ('2025-02-26', 'Data Analytics Roadmap', 'Lena Vogel, James Wu, Ryan Chow', '60 min',
         'Prioritize analytics dashboard features; evaluate BI tool upgrade options', 'Lena Vogel',
         'Current BI tool lacks cohort analysis and funnel visualization. Three alternatives evaluated: Tableau, Looker, and Metabase. Decision deferred to next month pending cost-benefit analysis. Ryan to prepare comparison matrix. Priority feature list compiled with 8 items for this quarter.'),
        ('2025-03-03', 'Q2 Budget Planning Kickoff', 'Karen Yuen, Sarah Chen, All VPs', '120 min',
         'Submit departmental Q2 budget requests by March 10; identify cost savings', 'Karen Yuen',
         'Q2 budget process officially kicked off. Total target budget set at $6.2M with 8% YoY growth allowance. Each department to submit detailed requests using new zero-based budgeting template. Karen to host office hours March 5 and 7 for questions. Hard deadline March 10.'),
        ('2025-03-05', 'Engineering Architecture Review', 'David Lee, Carlos Rivera, James Wu', '90 min',
         'Document microservices migration plan; assign technical leads per service', 'David Lee',
         'Monolith-to-microservices migration plan reviewed and approved in principle. Eight services identified for extraction in Q2 and Q3. Technical leads assigned to each. Risk: team capacity during migration. Contingency plan to hire two contractors if needed. ADR document to be drafted by March 12.'),
        ('2025-03-07', 'Customer Onboarding Process Review', 'Rachel Kim, Sofia Martini, Omar Hassan', '60 min',
         'Reduce onboarding time from 14 days to 7 days; update playbook', 'Rachel Kim',
         'Current 14-day onboarding analyzed step-by-step. Identified three bottlenecks: manual data import, delayed credentials setup, and scheduling friction. Automating credentials step expected to save 3 days alone. Revised playbook to be drafted by March 14. Pilot with next three customers.'),
        ('2025-03-10', 'Legal and Compliance Review', 'Tom Fischer, Natalie Brooks, Karen Yuen', '60 min',
         'Update vendor contracts for GDPR compliance; review DPA templates', 'Tom Fischer',
         'GDPR audit identified 12 vendor contracts requiring updated Data Processing Agreements. Tom to lead outreach to all vendors. Internal data retention policy also requires update. Natalie to review employee data handling procedures. All updates to be completed by April 30.'),
        ('2025-03-12', 'Content Marketing Strategy', 'Priya Sharma, Alicia Torres, Ryan Chow', '75 min',
         'Launch blog editorial calendar for Q2; identify three guest contributors', 'Priya Sharma',
         'Q2 editorial calendar drafted with 24 planned articles across SEO, thought leadership, and product updates. Guest contributor program launching in April — Ryan to identify candidates. Podcast series idea proposed; feasibility study assigned to Alicia. Budget for content creation: $15K for Q2.'),
        ('2025-03-14', 'Platform Reliability Review', 'Carlos Rivera, David Lee, James Wu', '60 min',
         'Implement SLO monitoring dashboard; review on-call rotation schedule', 'Carlos Rivera',
         'Platform availability last 30 days: 99.7%, slightly below 99.9% SLO target. Two incidents analyzed — both related to database connection pool exhaustion. Monitoring dashboard to be built in Grafana by March 21. On-call rotation revised to reduce consecutive weekend coverage. PagerDuty integration to be updated.'),
        ('2025-03-17', 'Quarterly All-Hands Prep', 'Sarah Chen, All Department Heads', '60 min',
         'Finalize all-hands agenda; coordinate slide submissions from each team', 'Sarah Chen',
         'Q1 All-Hands scheduled for March 28, 10am company-wide. Agenda confirmed: CEO update, financial highlights, product demo, team spotlight, and Q&A. Each department to submit slides by March 24. Sarah will do final review March 26. Town hall format with live Q&A via Slido.'),
    ]

    # --- Write data with explicit wrap_text=False and vertical=bottom for column G ---
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 7:  # Column G - Notes column
                # Explicitly set wrap=False and vertical alignment=bottom (pre-task state)
                cell.alignment = Alignment(wrap_text=False, vertical='bottom')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
