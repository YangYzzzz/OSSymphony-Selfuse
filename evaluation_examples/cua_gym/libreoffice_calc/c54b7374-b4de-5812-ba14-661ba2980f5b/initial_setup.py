"""
Initial Setup: Sales quota rep target with VLOOKUP lookup tables
Task ID: calc_sales_quota_rep_target_033
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_quota_rep_target_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: RepQuotas ---
    ws1 = wb.active
    ws1.title = 'RepQuotas'

    # Headers
    headers = ['Rep Name', 'Seniority Level', 'Territory', 'Base Quota',
               'Territory Multiplier', 'Final Quota', 'Rounded Quota']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # 25 realistic sales reps — columns D through G intentionally left empty
    reps = [
        ('Sarah Chen',       'Senior',    'Northeast'),
        ('Marcus Johnson',   'Mid',       'Southeast'),
        ('Emily Rodriguez',  'Junior',    'Midwest'),
        ('David Kim',        'Principal', 'West'),
        ('Jessica Patel',    'Senior',    'Southwest'),
        ('Brian O\'Connell', 'Mid',       'Northeast'),
        ('Amanda Torres',    'Junior',    'Southeast'),
        ('Kevin Williams',   'Principal', 'Midwest'),
        ('Rachel Nguyen',    'Senior',    'West'),
        ('Daniel Park',      'Mid',       'Southwest'),
        ('Lauren Mitchell',  'Junior',    'Northeast'),
        ('Tyler Brooks',     'Senior',    'Southeast'),
        ('Megan Foster',     'Principal', 'Midwest'),
        ('James Thornton',   'Mid',       'West'),
        ('Brittany Hayes',   'Junior',    'Southwest'),
        ('Nathan Cooper',    'Senior',    'Northeast'),
        ('Samantha Rivera',  'Mid',       'Southeast'),
        ('Christopher Lee',  'Junior',    'Midwest'),
        ('Ashley Morgan',    'Principal', 'West'),
        ('Robert Sullivan',  'Senior',    'Southwest'),
        ('Danielle Watson',  'Mid',       'Northeast'),
        ('Matthew Harris',   'Junior',    'Southeast'),
        ('Stephanie Clark',  'Senior',    'Midwest'),
        ('Joshua Bennett',   'Principal', 'West'),
        ('Kimberly Stewart', 'Mid',       'Southwest'),
    ]

    for r, (name, seniority, territory) in enumerate(reps, 2):
        ws1.cell(row=r, column=1, value=name)
        ws1.cell(row=r, column=2, value=seniority)
        ws1.cell(row=r, column=3, value=territory)
        # Columns D (4), E (5), F (6), G (7) left empty intentionally

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 14
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 16

    # --- Sheet 2: SeniorityTable ---
    ws2 = wb.create_sheet('SeniorityTable')

    ws2['A1'] = 'Level'
    ws2['B1'] = 'Base Quota'

    seniority_data = [
        ('Junior',    400000),
        ('Mid',       650000),
        ('Senior',    900000),
        ('Principal', 1200000),
    ]
    for r, (level, quota) in enumerate(seniority_data, 2):
        ws2.cell(row=r, column=1, value=level)
        ws2.cell(row=r, column=2, value=quota)

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 14

    # --- Sheet 3: TerritoryMultipliers ---
    ws3 = wb.create_sheet('TerritoryMultipliers')

    ws3['A1'] = 'Territory'
    ws3['B1'] = 'Multiplier'

    territory_data = [
        ('Northeast', 1.2),
        ('Southeast', 1.0),
        ('Midwest',   0.9),
        ('West',      1.4),
        ('Southwest', 0.8),
    ]
    for r, (territory, multiplier) in enumerate(territory_data, 2):
        ws3.cell(row=r, column=1, value=territory)
        ws3.cell(row=r, column=2, value=multiplier)

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  RepQuotas: 25 reps, columns D-G empty')
    print(f'  SeniorityTable: 4 levels')
    print(f'  TerritoryMultipliers: 5 territories')


create_initial()
