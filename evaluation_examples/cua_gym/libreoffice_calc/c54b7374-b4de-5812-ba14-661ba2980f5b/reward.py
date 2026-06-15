"""
Reward Script: Set individual rep quotas based on seniority and territory
Task ID: calc_sales_quota_rep_target_033
Domain: libreoffice_calc
Scoring:
  - Component 1: VLOOKUP Base Quota formulas in D2:D26 (0.25 pts)
  - Component 2: VLOOKUP Territory Multiplier formulas in E2:E26 (0.25 pts)
  - Component 3: Final Quota multiplication formulas in F2:F26 (0.20 pts)
  - Component 4: ROUND formula for Rounded Quota in G2:G26 (0.20 pts)
  - Component 5: Currency format applied to D, F, G columns (0.10 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_quota_rep_target_033'

def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, no spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: RepQuotas sheet must exist
    if 'RepQuotas' not in wb.sheetnames:
        print("FAIL: Sheet 'RepQuotas' not found")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['RepQuotas']

    # -------------------------------------------------------------------------
    # Component 1: VLOOKUP Base Quota formulas in D2:D26 (0.25 pts)
    # Task requires: =VLOOKUP(B{row}, SeniorityTable.$A:$B, 2, 0)
    # The initial file has None in D2:D26 — formulas must be ADDED.
    # -------------------------------------------------------------------------
    try:
        d_formula_count = 0
        d_formula_correct = 0
        for row in range(2, 27):
            val = ws.cell(row=row, column=4).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                d_formula_count += 1
                norm = normalize_formula(val)
                # Accept variations: VLOOKUP referencing B{row} and SeniorityTable column A:B, returning col 2
                # Exact expected (case-insensitive): =VLOOKUP(B{row},SeniorityTable.$A:$B,2,0)
                # Also accept without $ signs: =VLOOKUP(B{row},SeniorityTable.A:B,2,0) or !A:B notation
                if ('VLOOKUP' in norm and
                        f'B{row}' in norm and
                        'SENIORITYTABLE' in norm and
                        ',2,' in norm):
                    d_formula_correct += 1

        if d_formula_correct == 25:
            print(f"PASS: Component 1 — All 25 VLOOKUP Base Quota formulas present in D2:D26 (0.25 pts)")
            total_score += 0.25
        elif d_formula_correct >= 20:
            print(f"PARTIAL: Component 1 — {d_formula_correct}/25 VLOOKUP Base Quota formulas correct in D column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Only {d_formula_correct}/25 VLOOKUP Base Quota formulas found in D column "
                  f"({d_formula_count} formula cells total)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: VLOOKUP Territory Multiplier formulas in E2:E26 (0.25 pts)
    # Task requires: =VLOOKUP(C{row}, TerritoryMultipliers.$A:$B, 2, 0)
    # The initial file has None in E2:E26 — formulas must be ADDED.
    # -------------------------------------------------------------------------
    try:
        e_formula_count = 0
        e_formula_correct = 0
        for row in range(2, 27):
            val = ws.cell(row=row, column=5).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                e_formula_count += 1
                norm = normalize_formula(val)
                # Expected: =VLOOKUP(C{row},TerritoryMultipliers.$A:$B,2,0)
                if ('VLOOKUP' in norm and
                        f'C{row}' in norm and
                        'TERRITORYMULTIPLIERS' in norm and
                        ',2,' in norm):
                    e_formula_correct += 1

        if e_formula_correct == 25:
            print(f"PASS: Component 2 — All 25 VLOOKUP Territory Multiplier formulas present in E2:E26 (0.25 pts)")
            total_score += 0.25
        elif e_formula_correct >= 20:
            print(f"PARTIAL: Component 2 — {e_formula_correct}/25 VLOOKUP Territory Multiplier formulas correct in E column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Only {e_formula_correct}/25 VLOOKUP Territory Multiplier formulas found in E column "
                  f"({e_formula_count} formula cells total)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Final Quota multiplication formulas in F2:F26 (0.20 pts)
    # Task requires: =D{row}*E{row}
    # The initial file has None in F2:F26 — formulas must be ADDED.
    # -------------------------------------------------------------------------
    try:
        f_formula_count = 0
        f_formula_correct = 0
        for row in range(2, 27):
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                f_formula_count += 1
                norm = normalize_formula(val)
                # Expected: =D{row}*E{row} — product of Base Quota and Territory Multiplier
                if f'D{row}' in norm and f'E{row}' in norm and '*' in norm:
                    f_formula_correct += 1

        if f_formula_correct == 25:
            print(f"PASS: Component 3 — All 25 Final Quota multiplication formulas present in F2:F26 (0.20 pts)")
            total_score += 0.20
        elif f_formula_correct >= 20:
            print(f"PARTIAL: Component 3 — {f_formula_correct}/25 Final Quota formulas correct in F column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Only {f_formula_correct}/25 Final Quota formulas found in F column "
                  f"({f_formula_count} formula cells total)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: ROUND formula for Rounded Quota in G2:G26 (0.20 pts)
    # Task requires: =ROUND(F{row}/50000,0)*50000 — nearest $50K
    # The initial file has None in G2:G26 — formulas must be ADDED.
    # -------------------------------------------------------------------------
    try:
        g_formula_count = 0
        g_formula_correct = 0
        for row in range(2, 27):
            val = ws.cell(row=row, column=7).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                g_formula_count += 1
                norm = normalize_formula(val)
                # Expected: =ROUND(F{row}/50000,0)*50000
                # Must contain ROUND, reference F{row}, use 50000, and multiply back by 50000
                if ('ROUND' in norm and
                        f'F{row}' in norm and
                        '50000' in norm):
                    g_formula_correct += 1

        if g_formula_correct == 25:
            print(f"PASS: Component 4 — All 25 ROUND Quota formulas present in G2:G26 (0.20 pts)")
            total_score += 0.20
        elif g_formula_correct >= 20:
            print(f"PARTIAL: Component 4 — {g_formula_correct}/25 ROUND Quota formulas correct in G column (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — Only {g_formula_correct}/25 ROUND Quota formulas found in G column "
                  f"({g_formula_count} formula cells total)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Currency format on D, F, G columns (0.10 pts)
    # Task requires: Currency format on D, F, G columns (D=Base Quota, F=Final Quota, G=Rounded Quota)
    # Checks that at least 2 of these 3 columns use a currency-like number format.
    # The initial file has no number format on these cells (they are None/empty).
    # -------------------------------------------------------------------------
    try:
        currency_columns_formatted = 0
        currency_pattern = re.compile(r'[\$\#,]', re.IGNORECASE)

        for col_idx, col_name in [(4, 'D'), (6, 'F'), (7, 'G')]:
            # Check at least one representative cell for currency format
            formatted_count = 0
            for row in range(2, 27):
                cell = ws.cell(row=row, column=col_idx)
                fmt = cell.number_format or 'General'
                if currency_pattern.search(fmt):
                    formatted_count += 1
            if formatted_count >= 20:  # majority of cells should be formatted
                currency_columns_formatted += 1

        if currency_columns_formatted >= 3:
            print(f"PASS: Component 5 — Currency format applied to all 3 columns D, F, G (0.10 pts)")
            total_score += 0.10
        elif currency_columns_formatted >= 2:
            print(f"PASS: Component 5 — Currency format applied to {currency_columns_formatted}/3 columns (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Currency format only on {currency_columns_formatted}/3 columns (D, F, G)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
