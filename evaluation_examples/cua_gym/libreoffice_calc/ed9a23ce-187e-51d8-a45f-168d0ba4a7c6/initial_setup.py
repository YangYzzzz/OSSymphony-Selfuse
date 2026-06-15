"""
Initial Setup: Create department satisfaction spreadsheet with Overview sheet
Task ID: calc_mcp_044
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Shared styles ---
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    def style_header(ws, row, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def style_data_cell(ws, row, col):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # =============================================
    # Sheet 1: HR
    # =============================================
    ws_hr = wb.active
    ws_hr.title = "HR"
    hr_headers = ["Employee", "Role", "Satisfaction Score", "Years of Service", "Status"]
    for c, h in enumerate(hr_headers, 1):
        ws_hr.cell(row=1, column=c, value=h)
    style_header(ws_hr, 1, len(hr_headers))

    hr_data = [
        ["Elena Rodriguez", "HR Director", 82, 8, "Active"],
        ["James Park", "Recruiter", 75, 3, "Active"],
        ["Amara Okafor", "Benefits Coordinator", 88, 5, "Active"],
        ["Liam Foster", "HR Analyst", 78, 2, "Active"],   # C5 = 78
        ["Nina Vasquez", "Training Specialist", 84, 6, "Active"],
        ["David Kim", "Payroll Manager", 71, 4, "Active"],
        ["Sophie Laurent", "HR Generalist", 79, 1, "Active"],
        ["Marcus Brown", "Compliance Officer", 86, 7, "Active"],
    ]
    for r, row_data in enumerate(hr_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_hr.cell(row=r, column=c, value=val)
            style_data_cell(ws_hr, r, c)

    ws_hr.column_dimensions["A"].width = 22
    ws_hr.column_dimensions["B"].width = 24
    ws_hr.column_dimensions["C"].width = 20
    ws_hr.column_dimensions["D"].width = 18
    ws_hr.column_dimensions["E"].width = 12

    # =============================================
    # Sheet 2: Engineering
    # =============================================
    ws_eng = wb.create_sheet("Engineering")
    eng_headers = ["Employee", "Role", "Satisfaction Score", "Years of Service", "Status"]
    for c, h in enumerate(eng_headers, 1):
        ws_eng.cell(row=1, column=c, value=h)
    style_header(ws_eng, 1, len(eng_headers))

    eng_data = [
        ["Raj Patel", "VP Engineering", 90, 10, "Active"],
        ["Sarah Chen", "Senior Developer", 87, 6, "Active"],
        ["Tom Williams", "DevOps Lead", 82, 4, "Active"],
        ["Yuki Tanaka", "Software Engineer", 85, 3, "Active"],   # C5 = 85
        ["Alex Novak", "QA Lead", 79, 5, "Active"],
        ["Priya Sharma", "Frontend Developer", 91, 2, "Active"],
        ["Chris O'Brien", "Backend Developer", 83, 3, "Active"],
        ["Maya Johnson", "Data Engineer", 88, 4, "Active"],
        ["Leo Martinez", "Security Analyst", 76, 7, "Active"],
    ]
    for r, row_data in enumerate(eng_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_eng.cell(row=r, column=c, value=val)
            style_data_cell(ws_eng, r, c)

    ws_eng.column_dimensions["A"].width = 22
    ws_eng.column_dimensions["B"].width = 24
    ws_eng.column_dimensions["C"].width = 20
    ws_eng.column_dimensions["D"].width = 18
    ws_eng.column_dimensions["E"].width = 12

    # =============================================
    # Sheet 3: Marketing
    # =============================================
    ws_mkt = wb.create_sheet("Marketing")
    mkt_headers = ["Employee", "Role", "Satisfaction Score", "Years of Service", "Status"]
    for c, h in enumerate(mkt_headers, 1):
        ws_mkt.cell(row=1, column=c, value=h)
    style_header(ws_mkt, 1, len(mkt_headers))

    mkt_data = [
        ["Isabella Torres", "Marketing Director", 80, 9, "Active"],
        ["Ryan Cooper", "Content Strategist", 74, 3, "Active"],
        ["Hannah Lee", "Social Media Manager", 81, 2, "Active"],
        ["Oliver Grant", "Brand Analyst", 72, 4, "Active"],   # C5 = 72
        ["Zara Ahmed", "Digital Marketing Lead", 85, 5, "Active"],
        ["Ethan Brooks", "SEO Specialist", 77, 1, "Active"],
        ["Chloe Martin", "Campaign Manager", 83, 6, "Active"],
    ]
    for r, row_data in enumerate(mkt_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_mkt.cell(row=r, column=c, value=val)
            style_data_cell(ws_mkt, r, c)

    ws_mkt.column_dimensions["A"].width = 22
    ws_mkt.column_dimensions["B"].width = 24
    ws_mkt.column_dimensions["C"].width = 20
    ws_mkt.column_dimensions["D"].width = 18
    ws_mkt.column_dimensions["E"].width = 12

    # =============================================
    # Sheet 4: Overview (first sheet in order)
    # =============================================
    ws_ov = wb.create_sheet("Overview", 0)  # insert at position 0

    # Title row
    ws_ov.cell(row=1, column=1, value="Metric")
    ws_ov.cell(row=1, column=2, value="Value")
    ws_ov.cell(row=1, column=3, value="Notes")
    style_header(ws_ov, 1, 3)

    # Row 2: Average Satisfaction - B2 must be EMPTY (task target)
    ws_ov.cell(row=2, column=1, value="Avg Department Satisfaction")
    # B2 intentionally left empty - this is the task target
    ws_ov.cell(row=2, column=3, value="Average of C5 across HR, Engineering, Marketing")

    # Additional overview rows for realism
    ws_ov.cell(row=3, column=1, value="Total Headcount")
    ws_ov.cell(row=3, column=2, value=24)
    ws_ov.cell(row=3, column=3, value="Sum of all department employees")

    ws_ov.cell(row=4, column=1, value="Departments Tracked")
    ws_ov.cell(row=4, column=2, value=3)
    ws_ov.cell(row=4, column=3, value="HR, Engineering, Marketing")

    ws_ov.cell(row=5, column=1, value="Report Period")
    ws_ov.cell(row=5, column=2, value="Q1 2026")
    ws_ov.cell(row=5, column=3, value="January - March 2026")

    ws_ov.cell(row=6, column=1, value="Last Updated")
    ws_ov.cell(row=6, column=2, value="2026-03-31")

    for r in range(2, 7):
        for c in range(1, 4):
            style_data_cell(ws_ov, r, c)

    ws_ov.column_dimensions["A"].width = 30
    ws_ov.column_dimensions["B"].width = 18
    ws_ov.column_dimensions["C"].width = 45

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
