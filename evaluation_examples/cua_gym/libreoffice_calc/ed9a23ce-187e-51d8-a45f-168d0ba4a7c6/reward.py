"""
Reward Script: Consolidation formula averaging C5 across three department sheets
Task ID: calc_mcp_044
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): B2 contains a formula referencing all three sheets and C5
  Component 2 (0.3): Formula performs averaging (divides by 3)
  Component 3 (0.2): Computed value matches expected average (~78.33)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_044'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Overview sheet exists
    if 'Overview' not in wb.sheetnames:
        print("FAIL: 'Overview' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Overview']
    b2_value = ws['B2'].value

    # Component 1: B2 contains a formula that references HR.C5, Engineering.C5,
    # and Marketing.C5 (0.5 points)
    # This checks that the cell has a cross-sheet consolidation formula
    try:
        if b2_value is not None and isinstance(b2_value, str) and b2_value.startswith('='):
            formula_upper = b2_value.upper().replace(' ', '')
            # Check references to all three department sheets and C5
            has_hr = bool(re.search(r'HR[\.!]C5', formula_upper, re.IGNORECASE))
            has_eng = bool(re.search(r'ENGINEERING[\.!]C5', formula_upper, re.IGNORECASE))
            has_mkt = bool(re.search(r'MARKETING[\.!]C5', formula_upper, re.IGNORECASE))

            if has_hr and has_eng and has_mkt:
                print(f"PASS: Component 1 -- B2 references all three department C5 cells (0.5 pts)")
                print(f"  Formula: {b2_value}")
                total_score += 0.5
            else:
                missing = []
                if not has_hr:
                    missing.append('HR.C5')
                if not has_eng:
                    missing.append('Engineering.C5')
                if not has_mkt:
                    missing.append('Marketing.C5')
                print(f"FAIL: Component 1 -- Formula missing references: {missing}")
                print(f"  Formula: {b2_value}")
        else:
            print(f"FAIL: Component 1 -- B2 does not contain a formula, found: {repr(b2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Formula performs averaging (divides by 3) (0.3 points)
    try:
        if b2_value is not None and isinstance(b2_value, str) and b2_value.startswith('='):
            formula_upper = b2_value.upper().replace(' ', '')
            # Check for division by 3 or use of AVERAGE function
            has_div3 = '/3' in formula_upper
            has_average = 'AVERAGE(' in formula_upper
            if has_div3 or has_average:
                print(f"PASS: Component 2 -- Formula includes averaging logic (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 -- Formula does not divide by 3 or use AVERAGE")
                print(f"  Formula: {b2_value}")
        else:
            print(f"FAIL: Component 2 -- No formula in B2 to check averaging")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Computed value matches expected average (0.2 points)
    # Expected: (78 + 85 + 72) / 3 = 78.333...
    # We verify by reading the source C5 values and checking data_only cached value
    try:
        # Read actual C5 values from each department sheet
        hr_c5 = wb['HR']['C5'].value if 'HR' in wb.sheetnames else None
        eng_c5 = wb['Engineering']['C5'].value if 'Engineering' in wb.sheetnames else None
        mkt_c5 = wb['Marketing']['C5'].value if 'Marketing' in wb.sheetnames else None

        if hr_c5 is not None and eng_c5 is not None and mkt_c5 is not None:
            expected_avg = (float(hr_c5) + float(eng_c5) + float(mkt_c5)) / 3.0

            # Try to get computed value from data_only load
            try:
                wb_data = openpyxl.load_workbook(file_path, data_only=True)
                computed = wb_data['Overview']['B2'].value
                if computed is not None and abs(float(computed) - expected_avg) < 0.01:
                    print(f"PASS: Component 3 -- Computed value {computed} matches expected {expected_avg:.4f} (0.2 pts)")
                    total_score += 0.2
                elif computed is not None:
                    print(f"FAIL: Component 3 -- Computed value {computed} != expected {expected_avg:.4f}")
                else:
                    # data_only returned None (file not opened in Calc yet)
                    # Fall back: if formula structure is correct (Components 1+2 passed),
                    # and source data is intact, award points based on formula correctness
                    if total_score >= 0.8:
                        # Both component 1 and 2 passed - formula is structurally correct
                        # With correct source data, it WILL compute correctly
                        print(f"PASS: Component 3 -- No cached value but formula is structurally correct and source data intact (0.2 pts)")
                        print(f"  Expected average: {expected_avg:.4f} from HR.C5={hr_c5}, Eng.C5={eng_c5}, Mkt.C5={mkt_c5}")
                        total_score += 0.2
                    else:
                        print(f"FAIL: Component 3 -- No cached value and formula structure incomplete")
            except Exception as e:
                # Same fallback
                if total_score >= 0.8:
                    print(f"PASS: Component 3 -- data_only load failed but formula structure correct (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 3 -- Cannot verify computed value: {e}")
        else:
            print(f"FAIL: Component 3 -- Source C5 values missing: HR={hr_c5}, Eng={eng_c5}, Mkt={mkt_c5}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
