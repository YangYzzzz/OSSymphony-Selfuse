"""
Reward Script: Weekly Attendance Trend Chart
Task ID: calc_edu_attendance_trend_038
Domain: libreoffice_calc
Scoring:
  Component 1: Weekly Avg formulas in G2:G19 (AVERAGE pattern)         — 0.35 pts
  Component 2: Semester Avg formulas in H2:H19 (absolute ref pattern)  — 0.25 pts
  Component 3: Line chart with exactly 2 series on sheet               — 0.25 pts
  Component 4: Chart title is 'Weekly Attendance Trend'                 — 0.15 pts
Total: 1.0
"""

import os
import openpyxl
from openpyxl.chart import LineChart

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_attendance_trend_038'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet 'AttendanceTrend' must exist
    if 'AttendanceTrend' not in wb.sheetnames:
        print("FAIL: Sheet 'AttendanceTrend' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['AttendanceTrend']

    # Component 1: Weekly Avg AVERAGE formulas in G2:G19 (0.35 points)
    # Each row should have =AVERAGE(B{n}:F{n}) — checking that the formula pattern
    # is consistent with row-by-row AVERAGE of Monday-Friday columns
    try:
        g_formula_count = 0
        g_formula_pattern_ok = 0
        for row in range(2, 20):  # rows 2-19 (18 weeks)
            val = ws.cell(row=row, column=7).value  # column G
            if val is not None and isinstance(val, str) and val.upper().startswith('=AVERAGE('):
                g_formula_count += 1
                # Check the formula references the correct row: B{row}:F{row}
                expected_pattern = f'=AVERAGE(B{row}:F{row})'
                if val.upper().replace(' ', '') == expected_pattern.upper().replace(' ', ''):
                    g_formula_pattern_ok += 1

        if g_formula_count == 18 and g_formula_pattern_ok == 18:
            print(f"PASS: Component 1 — All 18 weekly avg formulas correct (=AVERAGE(B2:F2) pattern) (0.35 pts)")
            total_score += 0.35
        elif g_formula_count >= 15:
            # Partial credit: most rows have correct AVERAGE formula
            partial = round(0.35 * g_formula_count / 18, 4)
            print(f"PARTIAL: Component 1 — {g_formula_count}/18 weekly avg formulas found ({g_formula_pattern_ok} with correct row refs) — awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {g_formula_count}/18 weekly avg formulas found in column G (expected =AVERAGE(B:F) pattern)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Semester Avg formulas in H2:H19 with absolute reference (0.25 points)
    # All rows should have =AVERAGE($G$2:$G$19) — absolute reference so same value in all rows
    try:
        h_formula_count = 0
        expected_h_formula = '=AVERAGE($G$2:$G$19)'
        for row in range(2, 20):  # rows 2-19 (18 weeks)
            val = ws.cell(row=row, column=8).value  # column H
            if val is not None and isinstance(val, str):
                normalized_val = val.upper().replace(' ', '')
                normalized_expected = expected_h_formula.upper().replace(' ', '')
                if normalized_val == normalized_expected:
                    h_formula_count += 1

        if h_formula_count == 18:
            print(f"PASS: Component 2 — All 18 semester avg formulas correct (=AVERAGE($G$2:$G$19) with absolute ref) (0.25 pts)")
            total_score += 0.25
        elif h_formula_count > 0:
            # Also accept a non-absolute but equivalent formula
            # Re-check with non-absolute version
            h_formula_count_alt = 0
            for row in range(2, 20):
                val = ws.cell(row=row, column=8).value
                if val is not None and isinstance(val, str) and 'AVERAGE' in val.upper() and 'G' in val.upper():
                    h_formula_count_alt += 1
            if h_formula_count_alt == 18:
                print(f"PARTIAL: Component 2 — {h_formula_count}/18 rows have exact absolute-reference semester avg formula; {h_formula_count_alt} have some AVERAGE of G formula — awarding 0.15 pts")
                total_score += 0.15
            else:
                partial = round(0.25 * h_formula_count / 18, 4)
                print(f"PARTIAL: Component 2 — {h_formula_count}/18 semester avg formulas correct — awarding {partial} pts")
                total_score += partial
        else:
            print(f"FAIL: Component 2 — No semester avg formulas found in column H (expected =AVERAGE($G$2:$G$19))")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Line chart with exactly 2 series on the sheet (0.25 points)
    # Task requires: a line chart showing 'Weekly Avg' and 'Semester Avg' series
    try:
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, LineChart)]

        if len(line_charts) >= 1:
            chart = line_charts[0]
            n_series = len(chart.series)
            if n_series == 2:
                print(f"PASS: Component 3 — Line chart found with exactly 2 series (Weekly Avg + Semester Avg) (0.25 pts)")
                total_score += 0.25
            elif n_series >= 1:
                print(f"PARTIAL: Component 3 — Line chart found but has {n_series} series (expected 2) — awarding 0.12 pts")
                total_score += 0.12
            else:
                print(f"FAIL: Component 3 — Line chart found but has no series")
        elif len(charts) >= 1:
            # Some chart exists but it's not a LineChart
            print(f"FAIL: Component 3 — Chart found but it is not a LineChart (got {type(charts[0]).__name__})")
        else:
            print(f"FAIL: Component 3 — No chart found on sheet 'AttendanceTrend'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Chart title is 'Weekly Attendance Trend' (0.15 points)
    try:
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, LineChart)]

        if len(line_charts) >= 1:
            chart = line_charts[0]
            chart_title = None
            try:
                # Try the nested rich text structure
                chart_title = chart.title.tx.rich.p[0].r[0].t
            except Exception:
                pass

            if chart_title is None:
                try:
                    chart_title = str(chart.title)
                except Exception:
                    pass

            if chart_title is not None:
                expected_title = 'Weekly Attendance Trend'
                if chart_title.strip() == expected_title:
                    print(f"PASS: Component 4 — Chart title is '{chart_title}' (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 4 — Chart title is '{chart_title}', expected '{expected_title}'")
            else:
                print(f"FAIL: Component 4 — Could not extract chart title")
        else:
            print(f"FAIL: Component 4 — No line chart found to check title")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
