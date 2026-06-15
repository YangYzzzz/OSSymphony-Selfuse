"""
Initial Setup: Weekly attendance trend spreadsheet - pre-task state
Task ID: calc_edu_attendance_trend_038
Domain: libreoffice_calc

Creates AttendanceTrend sheet with 18 weeks of daily attendance data.
Columns G (Weekly Avg) and H (Semester Avg) are LEFT EMPTY for the agent to fill.
No chart exists in initial file.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_attendance_trend_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'AttendanceTrend'

    # --- Row 1: Headers ---
    headers = ['Week', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Weekly Avg', 'Semester Avg']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- J1: Total class size ---
    ws['J1'] = 35
    ws['I1'] = 'Class Size'
    ws['I1'].font = Font(bold=True)

    # --- Rows 2-19: 18 weeks of realistic daily attendance data ---
    # Total class size = 35; attendance varies realistically week by week
    weekly_data = [
        # Week,  Mon, Tue, Wed, Thu, Fri
        ('Week 1',  32, 30, 31, 33, 28),
        ('Week 2',  31, 29, 30, 32, 27),
        ('Week 3',  33, 31, 32, 34, 29),
        ('Week 4',  30, 28, 29, 31, 26),
        ('Week 5',  34, 32, 33, 35, 30),
        ('Week 6',  29, 27, 28, 30, 25),
        ('Week 7',  31, 30, 32, 33, 28),
        ('Week 8',  28, 26, 27, 29, 24),   # mid-semester dip
        ('Week 9',  30, 29, 31, 32, 27),
        ('Week 10', 32, 31, 33, 34, 29),
        ('Week 11', 33, 32, 34, 35, 30),
        ('Week 12', 31, 30, 32, 33, 28),
        ('Week 13', 34, 33, 35, 35, 31),
        ('Week 14', 30, 29, 31, 32, 27),
        ('Week 15', 35, 34, 35, 35, 32),   # attendance picks up near end
        ('Week 16', 33, 32, 34, 35, 30),
        ('Week 17', 34, 33, 35, 35, 31),
        ('Week 18', 35, 35, 35, 35, 33),   # finals week - high attendance
    ]

    for i, (week, mon, tue, wed, thu, fri) in enumerate(weekly_data, 2):
        ws.cell(row=i, column=1, value=week)
        ws.cell(row=i, column=2, value=mon)
        ws.cell(row=i, column=3, value=tue)
        ws.cell(row=i, column=4, value=wed)
        ws.cell(row=i, column=5, value=thu)
        ws.cell(row=i, column=6, value=fri)
        # Columns G and H (7, 8) are intentionally left empty — agent must fill these

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 8
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 8

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: AttendanceTrend')
    print(f'  Rows: 18 weeks of daily attendance data (rows 2-19)')
    print(f'  Columns G and H are empty (agent must calculate weekly/semester averages)')
    print(f'  No chart present')


create_initial()
