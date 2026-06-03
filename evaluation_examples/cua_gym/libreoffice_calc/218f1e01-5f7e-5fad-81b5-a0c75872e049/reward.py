"""
Reward Script: Add polynomial (order 2) trend line to Temperature series with R-squared display
Task ID: calc_gg2_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Trendline exists on the Temperature series
  Component 2 (0.30): Trendline type is polynomial with order 2
  Component 3 (0.40): R-squared value display is enabled (dispRSqr=True)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_010'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Climate Data' sheet must exist
    if 'Climate Data' not in wb.sheetnames:
        print(f"FAIL: 'Climate Data' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Climate Data']

    # Precondition: Chart must exist
    if len(ws._charts) == 0:
        print("FAIL: No charts found on 'Climate Data' sheet")
        print("REWARD: 0.0")
        return 0.0

    chart = ws._charts[0]

    # Precondition: Chart must have at least one series
    if len(chart.series) == 0:
        print("FAIL: Chart has no data series")
        print("REWARD: 0.0")
        return 0.0

    # Find the Temperature series (should be the first/only series)
    temp_series = None
    for s in chart.series:
        temp_series = s
        break

    if temp_series is None:
        print("FAIL: No series found in chart")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Trendline exists on the Temperature series (0.30 points)
    try:
        trendline = temp_series.trendline
        if trendline is not None:
            print(f"PASS: Component 1 — Trendline exists on Temperature series (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — No trendline found on Temperature series")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Trendline is polynomial with order 2 (0.30 points)
    try:
        trendline = temp_series.trendline
        if trendline is not None:
            tl_type = trendline.trendlineType
            tl_order = trendline.order
            if tl_type == 'poly' and tl_order == 2:
                print(f"PASS: Component 2 — Trendline is polynomial order 2 (type={tl_type}, order={tl_order}) (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 2 — Expected poly order 2, found type={tl_type}, order={tl_order}")
        else:
            print(f"FAIL: Component 2 — No trendline to check type/order")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: R-squared value display is enabled (0.40 points)
    try:
        trendline = temp_series.trendline
        if trendline is not None:
            disp_rsqr = trendline.dispRSqr
            if disp_rsqr is True:
                print(f"PASS: Component 3 — R-squared display enabled (dispRSqr={disp_rsqr}) (0.40 pts)")
                total_score += 0.40
            else:
                print(f"FAIL: Component 3 — R-squared display not enabled (dispRSqr={disp_rsqr})")
        else:
            print(f"FAIL: Component 3 — No trendline to check R-squared display")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
