"""
Initial Setup: Create climate data spreadsheet with embedded line chart (no trend line)
Task ID: calc_gg2_010
Domain: libreoffice_calc
"""

import math
import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Climate Data sheet ---
    ws = wb.active
    ws.title = 'Climate Data'

    # Headers
    headers = ['Day', 'Date', 'Temperature']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Generate 30 days of temperature data with a curved seasonal pattern
    # Use a quadratic curve: T = -0.08*(day-15)^2 + 28 + small noise
    import random
    random.seed(42)

    base_date_year = 2025
    base_date_month = 6

    for day in range(1, 31):
        row = day + 1
        # Day number
        ws.cell(row=row, column=1, value=day)
        # Date
        date_str = f'2025-06-{day:02d}'
        ws.cell(row=row, column=2, value=date_str)
        ws.cell(row=row, column=2).number_format = 'yyyy-mm-dd'
        # Temperature with quadratic pattern + noise
        temp = -0.08 * (day - 15) ** 2 + 28 + random.uniform(-1.5, 1.5)
        temp = round(temp, 1)
        ws.cell(row=row, column=3, value=temp)

    # Set column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 15

    # Create embedded line chart
    chart = LineChart()
    chart.title = 'Daily Temperature Readings - June 2025'
    chart.y_axis.title = 'Temperature (\u00b0C)'
    chart.x_axis.title = 'Day'
    chart.style = 10
    chart.width = 20
    chart.height = 12

    # Temperature data series
    data_ref = Reference(ws, min_col=3, min_row=1, max_row=31)
    cats_ref = Reference(ws, min_col=1, min_row=2, max_row=31)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Style the series
    series = chart.series[0]
    series.graphicalProperties.line.width = 22000  # ~2pt

    # Place chart on the sheet starting at E2
    ws.add_chart(chart, 'E2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
