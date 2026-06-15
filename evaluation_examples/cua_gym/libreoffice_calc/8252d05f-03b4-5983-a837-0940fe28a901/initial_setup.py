"""
Initial Setup: Retail transaction data with two sheets — transactions in Sheet1, empty Sheet2.
Task ID: osworld_calc_pivot_dual_dimensions_006
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_pivot_dual_dimensions_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Transactions ---
    ws1 = wb.active
    ws1.title = 'Transactions'

    # Headers
    headers = ['Transaction ID', 'Store Location', 'Product Department', 'Units Sold', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic retail transaction data
    # Store locations: Downtown, Westside, Northgate, Eastfield, Southpark
    # Product departments: Electronics, Clothing, Home & Garden, Food & Beverage, Sports, Beauty
    data = [
        ['TXN-001', 'Downtown',   'Electronics',      12,   1440.00],
        ['TXN-002', 'Westside',   'Clothing',          8,    320.00],
        ['TXN-003', 'Northgate',  'Food & Beverage',  35,    875.00],
        ['TXN-004', 'Eastfield',  'Sports',            5,    450.00],
        ['TXN-005', 'Downtown',   'Clothing',         14,    560.00],
        ['TXN-006', 'Southpark',  'Home & Garden',    10,    750.00],
        ['TXN-007', 'Westside',   'Electronics',       7,    980.00],
        ['TXN-008', 'Northgate',  'Beauty',           20,    600.00],
        ['TXN-009', 'Eastfield',  'Food & Beverage',  50,   1250.00],
        ['TXN-010', 'Downtown',   'Home & Garden',     9,    675.00],
        ['TXN-011', 'Southpark',  'Electronics',      11,   1320.00],
        ['TXN-012', 'Westside',   'Sports',            6,    540.00],
        ['TXN-013', 'Northgate',  'Clothing',         18,    720.00],
        ['TXN-014', 'Eastfield',  'Beauty',           15,    450.00],
        ['TXN-015', 'Downtown',   'Food & Beverage',  40,   1000.00],
        ['TXN-016', 'Southpark',  'Clothing',         12,    480.00],
        ['TXN-017', 'Westside',   'Home & Garden',     8,    600.00],
        ['TXN-018', 'Northgate',  'Electronics',       9,   1080.00],
        ['TXN-019', 'Eastfield',  'Sports',            7,    630.00],
        ['TXN-020', 'Downtown',   'Beauty',           22,    660.00],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 18
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 14

    # --- Sheet 2: Analysis (empty — agent must populate with pivot summaries) ---
    ws2 = wb.create_sheet('Analysis')
    # Sheet2 is intentionally empty; the agent's task is to add the two pivot summaries

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
