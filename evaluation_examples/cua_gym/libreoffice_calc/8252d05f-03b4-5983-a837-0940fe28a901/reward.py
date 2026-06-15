"""
Reward Script: Dual pivot tables in Analysis sheet — Units Sold by Product Department
              and Revenue by Store Location.
Task ID: osworld_calc_pivot_dual_dimensions_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): Analysis sheet has data (at least one pivot table header present)
  Component 2 (0.3): First pivot table present — Units Sold by Product Department
                     with correct column headers and all 6 department rows with correct sums
  Component 3 (0.3): Second pivot table present — Revenue by Store Location
                     with correct column headers and all 5 store rows with correct sums
  Component 4 (0.2): Spot-check aggregated values (top-3 exact matches)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_dual_dimensions_006'

# Expected pivot table 1: Units Sold by Product Department
EXPECTED_DEPT_UNITS = {
    'Beauty': 57,
    'Clothing': 52,
    'Electronics': 39,
    'Food & Beverage': 125,
    'Home & Garden': 27,
    'Sports': 18,
}

# Expected pivot table 2: Revenue by Store Location
EXPECTED_STORE_REVENUE = {
    'Downtown': 4335,
    'Eastfield': 2780,
    'Northgate': 3275,
    'Southpark': 2550,
    'Westside': 2440,
}


def find_pivot_block(ws, header_keyword, search_max_row=30):
    """
    Scan the worksheet for a block containing a header keyword.
    Returns (title_row, data_start_row) or (None, None) if not found.
    """
    for row in ws.iter_rows(min_row=1, max_row=search_max_row, min_col=1, max_col=2):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and header_keyword.lower() in cell.value.lower():
                return cell.row, cell.row + 1  # data header follows the title row
    return None, None


def extract_pivot_data(ws, col_header_row, key_col, val_col, max_rows=20):
    """
    Starting from col_header_row + 1, read (key, value) pairs until an empty key cell.
    Returns a dict of {str: numeric}.
    """
    result = {}
    for r in range(col_header_row + 1, col_header_row + max_rows):
        key_cell = ws.cell(row=r, column=key_col)
        val_cell = ws.cell(row=r, column=val_col)
        if key_cell.value is None:
            break
        key = str(key_cell.value).strip()
        try:
            val = float(val_cell.value)
            result[key] = val
        except (TypeError, ValueError):
            result[key] = val_cell.value
    return result


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns a float between 0.0 and 1.0.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Determine the target sheet — task says "Sheet2"; actual sheet is named "Analysis"
    # Try both names; the agent may use either.
    analysis_ws = None
    for candidate in ['Analysis', 'Sheet2', 'analysis', 'sheet2']:
        if candidate in wb.sheetnames:
            analysis_ws = wb[candidate]
            break

    # Fall back: look for any non-Transactions sheet with data
    if analysis_ws is None:
        for name in wb.sheetnames:
            ws_candidate = wb[name]
            if name.lower() not in ('transactions', 'sheet1') and ws_candidate.max_row > 1:
                analysis_ws = ws_candidate
                break

    if analysis_ws is None:
        print("FAIL: No suitable Analysis/Sheet2 sheet found in workbook.")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    print(f"INFO: Using sheet '{analysis_ws.title}' for verification")

    # Component 1: Analysis sheet has content — at least one pivot header keyword (0.2 points)
    try:
        dept_keyword_count = 0
        loc_keyword_count = 0
        for row in analysis_ws.iter_rows(min_row=1, max_row=analysis_ws.max_row, min_col=1, max_col=4):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    lv = cell.value.lower()
                    if 'product department' in lv or 'units sold' in lv or 'department' in lv:
                        dept_keyword_count += 1
                    if 'store location' in lv or 'revenue' in lv or 'location' in lv:
                        loc_keyword_count += 1

        has_dept_header = dept_keyword_count > 0
        has_loc_header = loc_keyword_count > 0

        if has_dept_header and has_loc_header:
            print("PASS: Component 1 — Analysis sheet contains both pivot table headers (0.2 pts)")
            total_score += 0.2
        elif has_dept_header or has_loc_header:
            print(f"PARTIAL: Component 1 — Only one pivot header found (dept={has_dept_header}, loc={has_loc_header})")
            # No partial score for this component — need both
        else:
            print("FAIL: Component 1 — Analysis sheet has no recognizable pivot table headers")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: First pivot — Units Sold by Product Department (0.3 points)
    # Check: correct column headers and all 6 departments with correct totals
    try:
        # Find the title row for the first pivot
        dept_title_row, dept_col_header_row = find_pivot_block(
            analysis_ws,
            'product department',
            search_max_row=30
        )
        if dept_title_row is None:
            # Also try locating by column header row
            dept_title_row, dept_col_header_row = find_pivot_block(
                analysis_ws,
                'units sold',
                search_max_row=30
            )

        if dept_title_row is None:
            print("FAIL: Component 2 — Cannot find 'Product Department' pivot table header")
        else:
            # Verify column headers on col_header_row
            col1 = analysis_ws.cell(row=dept_col_header_row, column=1).value
            col2 = analysis_ws.cell(row=dept_col_header_row, column=2).value
            header_ok = (
                col1 and 'department' in str(col1).lower() and
                col2 and 'units' in str(col2).lower()
            )

            # Extract actual data
            actual_dept_data = extract_pivot_data(analysis_ws, dept_col_header_row, 1, 2)

            # Compare against expected
            correct_depts = 0
            for dept, expected_units in EXPECTED_DEPT_UNITS.items():
                actual = actual_dept_data.get(dept)
                if actual is not None and abs(float(actual) - expected_units) < 0.5:
                    correct_depts += 1
                else:
                    print(f"  MISMATCH dept '{dept}': expected {expected_units}, got {actual}")

            dept_row_count = len(actual_dept_data)
            all_correct = (correct_depts == len(EXPECTED_DEPT_UNITS))

            if header_ok and all_correct and dept_row_count == len(EXPECTED_DEPT_UNITS):
                print(f"PASS: Component 2 — Units Sold by Product Dept table complete: "
                      f"headers OK, {correct_depts}/6 departments correct (0.3 pts)")
                total_score += 0.3
            elif all_correct:
                print(f"PASS: Component 2 — Data correct ({correct_depts}/6 depts), "
                      f"header_ok={header_ok} (0.3 pts)")
                total_score += 0.3
            elif correct_depts >= 4:
                print(f"PARTIAL: Component 2 — {correct_depts}/6 departments correct, "
                      f"header_ok={header_ok}")
                # No partial score — need all 6 for full component
            else:
                print(f"FAIL: Component 2 — Only {correct_depts}/6 departments correct, "
                      f"header_ok={header_ok}, rows found={dept_row_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Second pivot — Revenue by Store Location (0.3 points)
    # Check: correct column headers and all 5 store locations with correct totals
    try:
        # Find the title row for the second pivot
        loc_title_row, loc_col_header_row = find_pivot_block(
            analysis_ws,
            'store location',
            search_max_row=40
        )
        if loc_title_row is None:
            loc_title_row, loc_col_header_row = find_pivot_block(
                analysis_ws,
                'revenue by',
                search_max_row=40
            )

        if loc_title_row is None:
            print("FAIL: Component 3 — Cannot find 'Store Location' pivot table header")
        else:
            # Verify column headers on loc_col_header_row
            col1 = analysis_ws.cell(row=loc_col_header_row, column=1).value
            col2 = analysis_ws.cell(row=loc_col_header_row, column=2).value
            loc_header_ok = (
                col1 and 'location' in str(col1).lower() and
                col2 and 'revenue' in str(col2).lower()
            )

            # Extract actual data
            actual_loc_data = extract_pivot_data(analysis_ws, loc_col_header_row, 1, 2)

            # Compare against expected
            correct_locs = 0
            for loc, expected_rev in EXPECTED_STORE_REVENUE.items():
                actual = actual_loc_data.get(loc)
                if actual is not None and abs(float(actual) - expected_rev) < 0.5:
                    correct_locs += 1
                else:
                    print(f"  MISMATCH loc '{loc}': expected {expected_rev}, got {actual}")

            loc_row_count = len(actual_loc_data)
            all_locs_correct = (correct_locs == len(EXPECTED_STORE_REVENUE))

            if loc_header_ok and all_locs_correct and loc_row_count == len(EXPECTED_STORE_REVENUE):
                print(f"PASS: Component 3 — Revenue by Store Location table complete: "
                      f"headers OK, {correct_locs}/5 locations correct (0.3 pts)")
                total_score += 0.3
            elif all_locs_correct:
                print(f"PASS: Component 3 — Data correct ({correct_locs}/5 locs), "
                      f"header_ok={loc_header_ok} (0.3 pts)")
                total_score += 0.3
            elif correct_locs >= 3:
                print(f"PARTIAL: Component 3 — {correct_locs}/5 locations correct, "
                      f"header_ok={loc_header_ok}")
            else:
                print(f"FAIL: Component 3 — Only {correct_locs}/5 locations correct, "
                      f"header_ok={loc_header_ok}, rows found={loc_row_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Spot-check 3 specific aggregated values for exact correctness (0.2 points)
    # Electronics total units = 39, Downtown total revenue = 4335, Food & Beverage total units = 125
    try:
        comp4_checks = 0

        # Check Electronics units sold (should be 39)
        dept_title_row2, dept_header_row2 = find_pivot_block(analysis_ws, 'product department', 30)
        if dept_title_row2 is None:
            dept_title_row2, dept_header_row2 = find_pivot_block(analysis_ws, 'units sold', 30)

        if dept_header_row2:
            dept_data_check = extract_pivot_data(analysis_ws, dept_header_row2, 1, 2)
            electronics_val = dept_data_check.get('Electronics')
            if electronics_val is not None and abs(float(electronics_val) - 39) < 0.5:
                comp4_checks += 1
                print(f"  CHECK: Electronics units = {electronics_val} (expected 39) ✓")
            else:
                print(f"  CHECK FAIL: Electronics units = {electronics_val} (expected 39)")

            food_val = dept_data_check.get('Food & Beverage')
            if food_val is not None and abs(float(food_val) - 125) < 0.5:
                comp4_checks += 1
                print(f"  CHECK: Food & Beverage units = {food_val} (expected 125) ✓")
            else:
                print(f"  CHECK FAIL: Food & Beverage units = {food_val} (expected 125)")

        loc_title_row2, loc_header_row2 = find_pivot_block(analysis_ws, 'store location', 40)
        if loc_header_row2:
            loc_data_check = extract_pivot_data(analysis_ws, loc_header_row2, 1, 2)
            downtown_val = loc_data_check.get('Downtown')
            if downtown_val is not None and abs(float(downtown_val) - 4335) < 0.5:
                comp4_checks += 1
                print(f"  CHECK: Downtown revenue = {downtown_val} (expected 4335) ✓")
            else:
                print(f"  CHECK FAIL: Downtown revenue = {downtown_val} (expected 4335)")

        if comp4_checks == 3:
            print(f"PASS: Component 4 — All 3 spot-check values correct (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Only {comp4_checks}/3 spot-checks passed")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
