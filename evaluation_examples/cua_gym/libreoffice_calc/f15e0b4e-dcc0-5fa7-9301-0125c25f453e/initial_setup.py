"""
Initial Setup: Create a spreadsheet with monthly sales detail data for pivot table task.
Task ID: calc_gcp_075
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def generate_date_in_month(year, month):
    """Generate a random date within a given month."""
    if month == 12:
        max_day = 31
    else:
        next_month = date(year, month + 1, 1)
        max_day = (next_month - timedelta(days=1)).day
    day = random.randint(1, max_day)
    return date(year, month, day)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MonthlySalesDetail'

    # --- Headers ---
    headers = ['TransID', 'SaleDate', 'Product', 'Quantity', 'Revenue']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    # --- Products with base revenue ranges and seasonal patterns ---
    products = ['Widget-A', 'Widget-B', 'Widget-C', 'Widget-D', 'Widget-E']

    # Base quantity and price per product
    product_config = {
        'Widget-A': {'qty_range': (5, 25), 'price_range': (45.00, 85.00)},
        'Widget-B': {'qty_range': (3, 18), 'price_range': (120.00, 250.00)},
        'Widget-C': {'qty_range': (10, 40), 'price_range': (15.00, 35.00)},
        'Widget-D': {'qty_range': (2, 12), 'price_range': (300.00, 550.00)},
        'Widget-E': {'qty_range': (8, 30), 'price_range': (60.00, 110.00)},
    }

    # Seasonal multipliers per month (1-indexed)
    seasonal = {
        1: 0.75, 2: 0.80, 3: 0.90, 4: 1.00, 5: 1.05,
        6: 1.10, 7: 1.00, 8: 0.95, 9: 1.05, 10: 1.15,
        11: 1.30, 12: 1.40,
    }

    # Product-specific seasonal modifiers
    product_seasonal_mod = {
        'Widget-A': {6: 1.3, 7: 1.4, 8: 1.3},   # summer boost
        'Widget-B': {11: 1.5, 12: 1.6},            # holiday boost
        'Widget-C': {1: 1.2, 2: 1.3, 9: 1.2},     # back-to-school/new-year
        'Widget-D': {3: 1.3, 4: 1.4, 10: 1.2},    # spring/Q4 boost
        'Widget-E': {5: 1.2, 6: 1.3, 11: 1.1},    # mid-year boost
    }

    # Generate 600 transactions spread across 12 months (50 per month)
    data_rows = []
    trans_id = 1
    for month in range(1, 13):
        n_transactions = 50
        for _ in range(n_transactions):
            product = random.choice(products)
            cfg = product_config[product]
            sale_date = generate_date_in_month(2024, month)

            base_mult = seasonal[month]
            prod_mod = product_seasonal_mod.get(product, {}).get(month, 1.0)
            mult = base_mult * prod_mod

            qty = max(1, int(random.randint(*cfg['qty_range']) * mult))
            unit_price = round(random.uniform(*cfg['price_range']) * mult, 2)
            revenue = round(qty * unit_price, 2)

            data_rows.append([trans_id, sale_date, product, qty, revenue])
            trans_id += 1

    # Sort by date for natural ordering
    data_rows.sort(key=lambda r: r[1])

    # Re-number TransIDs after sorting
    for i, row in enumerate(data_rows):
        row[0] = i + 1

    # --- Write data ---
    date_format = 'yyyy-mm-dd'
    currency_format = '$#,##0.00'
    int_format = '0'

    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.number_format = int_format
            elif c == 2:
                cell.number_format = date_format
            elif c == 4:
                cell.number_format = int_format
            elif c == 5:
                cell.number_format = currency_format

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total rows: {len(data_rows)} data rows + 1 header')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
