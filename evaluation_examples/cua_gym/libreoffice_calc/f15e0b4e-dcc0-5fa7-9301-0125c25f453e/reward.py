"""
Reward Script: Pivot table showing month-over-month % change for each product's sales
Task ID: calc_gcp_075
Domain: libreoffice_calc
Scoring:
  Component 1: Pivot-like sheet with % change data exists (0.15)
  Component 2: Correct structure - 5 products x 12 months (0.20)
  Component 3: January column shows N/A or blank (no prior month) (0.10)
  Component 4: Percentage change values are mathematically correct (0.35)
  Component 5: Values formatted as percentages (0.10)
  Component 6: Title/header indicates % difference from previous (0.10)
"""

import os
import openpyxl
from datetime import datetime
from collections import defaultdict

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_075'

MONTH_NAMES = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

PRODUCTS = ['Widget-A', 'Widget-B', 'Widget-C', 'Widget-D', 'Widget-E']


def compute_expected_revenue(wb):
    """Compute monthly revenue per product from the source data sheet."""
    ws = wb['MonthlySalesDetail']
    revenue = defaultdict(lambda: defaultdict(float))
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=5, values_only=True):
        trans_id, sale_date, product, quantity, rev = row
        if product is None or rev is None or sale_date is None:
            continue
        if isinstance(sale_date, datetime):
            month_idx = sale_date.month  # 1-12
        else:
            continue
        revenue[str(product).strip()][month_idx] += float(rev)
    return revenue


def compute_expected_pct_change(revenue):
    """Compute expected % change month-over-month from revenue data."""
    pct_change = {}
    for product in PRODUCTS:
        pct_change[product] = {}
        for m in range(1, 13):
            if m == 1:
                pct_change[product][m] = None  # No prior month
            else:
                prev = revenue[product].get(m - 1, 0)
                curr = revenue[product].get(m, 0)
                if prev != 0:
                    pct_change[product][m] = (curr - prev) / prev
                else:
                    pct_change[product][m] = None
    return pct_change


def find_pivot_sheet(wb):
    """Find the sheet that contains the pivot table with % change data.
    Returns (sheet, header_row, data_start_row, product_col, month_cols) or None.
    """
    for name in wb.sheetnames:
        if name == 'MonthlySalesDetail':
            continue
        ws = wb[name]
        # Look for a row with month names as headers
        for r in range(1, min(10, ws.max_row + 1)):
            row_vals = []
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                row_vals.append(str(v).strip() if v is not None else '')
            # Check if this row contains month names
            month_count = sum(1 for mv in MONTH_NAMES if mv in row_vals)
            if month_count >= 6:  # At least half the months present
                # Found header row
                header_row = r
                # Find product column (usually col A)
                product_col = None
                month_cols = {}
                for c_idx, val in enumerate(row_vals):
                    if val.lower() in ('product', 'products', ''):
                        if product_col is None:
                            product_col = c_idx + 1  # 1-based
                    for mi, mn in enumerate(MONTH_NAMES):
                        if val == mn:
                            month_cols[mi + 1] = c_idx + 1  # month 1-12 -> column
                if product_col is None:
                    product_col = 1  # Default to column A
                return ws, header_row, header_row + 1, product_col, month_cols
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: MonthlySalesDetail must exist
    if 'MonthlySalesDetail' not in wb.sheetnames:
        print("CRITICAL: MonthlySalesDetail sheet missing - cannot verify")
        print("REWARD: 0.0")
        return 0.0

    # Compute expected values from source data
    try:
        revenue = compute_expected_revenue(wb)
        expected_pct = compute_expected_pct_change(revenue)
    except Exception as e:
        print(f"CRITICAL: Cannot compute expected values: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot-like sheet with % change data exists (0.15 points)
    try:
        result = find_pivot_sheet(wb)
        if result is not None:
            pivot_ws, header_row, data_start, product_col, month_cols = result
            print(f"PASS: Component 1 - Found pivot sheet '{pivot_ws.title}' with month headers at row {header_row} (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 1 - No sheet found with product rows and month columns containing % change data")
            # Cannot proceed with further checks
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Component 2: Correct structure - 5 products x 12 months (0.20 points)
    try:
        # Find products in the data rows
        found_products = []
        product_rows = {}  # product_name -> row_number
        for r in range(data_start, data_start + 10):
            val = pivot_ws.cell(row=r, column=product_col).value
            if val is not None and str(val).strip() in PRODUCTS:
                p = str(val).strip()
                found_products.append(p)
                product_rows[p] = r

        products_ok = len(set(found_products)) == 5
        months_ok = len(month_cols) == 12

        if products_ok and months_ok:
            print(f"PASS: Component 2 - All 5 products and 12 months present (0.20 pts)")
            total_score += 0.20
        elif products_ok:
            print(f"PARTIAL: Component 2 - 5 products found but only {len(month_cols)} months (0.10 pts)")
            total_score += 0.10
        elif months_ok:
            print(f"PARTIAL: Component 2 - 12 months found but only {len(set(found_products))} products (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 - Found {len(set(found_products))} products and {len(month_cols)} months")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: January shows N/A or blank - no prior month (0.10 points)
    try:
        jan_col = month_cols.get(1)
        if jan_col is None:
            print("FAIL: Component 3 - January column not found")
        else:
            jan_na_count = 0
            for product, row_num in product_rows.items():
                val = pivot_ws.cell(row=row_num, column=jan_col).value
                if val is None or str(val).strip().upper() in ('N/A', 'NA', '-', '', '#N/A'):
                    jan_na_count += 1
            if jan_na_count == len(product_rows) and jan_na_count > 0:
                print(f"PASS: Component 3 - All {jan_na_count} products show N/A for January (0.10 pts)")
                total_score += 0.10
            elif jan_na_count > 0:
                print(f"PARTIAL: Component 3 - {jan_na_count}/{len(product_rows)} products show N/A for January (0.05 pts)")
                total_score += 0.05
            else:
                print("FAIL: Component 3 - January column has numeric values (should be N/A)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Percentage change values are mathematically correct (0.35 points)
    try:
        correct_count = 0
        total_checks = 0
        tolerance = 0.015  # Allow small rounding differences

        for product, row_num in product_rows.items():
            for month_num in range(2, 13):  # Feb-Dec
                col = month_cols.get(month_num)
                if col is None:
                    continue
                cell_val = pivot_ws.cell(row=row_num, column=col).value
                exp_val = expected_pct.get(product, {}).get(month_num)
                if exp_val is None:
                    continue
                total_checks += 1
                if cell_val is not None:
                    try:
                        actual = float(cell_val)
                        if abs(actual - exp_val) <= tolerance:
                            correct_count += 1
                        else:
                            print(f"  MISMATCH: {product} {MONTH_NAMES[month_num-1]}: actual={actual:.4f}, expected={exp_val:.4f}")
                    except (ValueError, TypeError):
                        print(f"  NON-NUMERIC: {product} {MONTH_NAMES[month_num-1]}: {cell_val}")

        if total_checks > 0:
            accuracy = correct_count / total_checks
            component_score = round(accuracy * 0.35, 4)
            if accuracy >= 0.95:
                print(f"PASS: Component 4 - {correct_count}/{total_checks} values correct ({accuracy:.1%}) (0.35 pts)")
                total_score += 0.35
            elif accuracy >= 0.5:
                print(f"PARTIAL: Component 4 - {correct_count}/{total_checks} values correct ({accuracy:.1%}) ({component_score} pts)")
                total_score += component_score
            else:
                print(f"FAIL: Component 4 - Only {correct_count}/{total_checks} values correct ({accuracy:.1%})")
        else:
            print("FAIL: Component 4 - No percentage values could be checked")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Values formatted as percentages (0.10 points)
    try:
        pct_formatted = 0
        checked_cells = 0
        for product, row_num in product_rows.items():
            for month_num in range(2, 13):
                col = month_cols.get(month_num)
                if col is None:
                    continue
                cell = pivot_ws.cell(row=row_num, column=col)
                if cell.value is not None:
                    checked_cells += 1
                    fmt = cell.number_format
                    if '%' in str(fmt):
                        pct_formatted += 1

        if checked_cells > 0:
            fmt_ratio = pct_formatted / checked_cells
            if fmt_ratio >= 0.8:
                print(f"PASS: Component 5 - {pct_formatted}/{checked_cells} cells formatted as percentage (0.10 pts)")
                total_score += 0.10
            elif fmt_ratio >= 0.3:
                print(f"PARTIAL: Component 5 - {pct_formatted}/{checked_cells} cells formatted as percentage (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 - Only {pct_formatted}/{checked_cells} cells formatted as percentage")
        else:
            print("FAIL: Component 5 - No data cells found to check formatting")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Title/header indicates % difference from previous (0.10 points)
    try:
        title_found = False
        # Check first few rows and sheet name for relevant keywords
        keywords = ['%', 'percent', 'difference', 'change', 'mom', 'month-over-month', 'previous']
        search_text = pivot_ws.title.lower()
        for r in range(1, min(data_start, 6)):
            for c in range(1, pivot_ws.max_column + 1):
                val = pivot_ws.cell(row=r, column=c).value
                if val is not None:
                    search_text += ' ' + str(val).lower()

        matched_keywords = [kw for kw in keywords if kw in search_text]
        if len(matched_keywords) >= 2:
            print(f"PASS: Component 6 - Title/header contains keywords: {matched_keywords} (0.10 pts)")
            total_score += 0.10
        elif len(matched_keywords) >= 1:
            print(f"PARTIAL: Component 6 - Title/header contains keyword: {matched_keywords} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 - No percentage/change keywords found in title or headers")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
