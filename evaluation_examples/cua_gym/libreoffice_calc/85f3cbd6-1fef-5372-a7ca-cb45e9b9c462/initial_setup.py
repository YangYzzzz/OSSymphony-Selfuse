"""
Initial Setup: Set all page margins to 1.5cm
Task ID: calc_gfl_049
Domain: libreoffice_calc

Creates a financial summary spreadsheet with default page margins (~2cm).
The agent must change all four margins to 1.5cm.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Summary ---
    ws = wb.active
    ws.title = 'Summary'

    # Headers
    headers = [
        'Category', 'Q1 Revenue', 'Q1 Expenses', 'Q1 Net',
        'Q2 Revenue', 'Q2 Expenses', 'Q2 Net', 'YTD Total'
    ]
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Financial data — 29 rows of data (rows 2-30)
    data = [
        ['Product Sales - Electronics', 145230, 89420, None, 162340, 95100, None, None],
        ['Product Sales - Furniture', 87650, 52180, None, 91200, 54600, None, None],
        ['Product Sales - Office Supplies', 34520, 18760, None, 38900, 21300, None, None],
        ['Service Revenue - Consulting', 210450, 125800, None, 228700, 132400, None, None],
        ['Service Revenue - Maintenance', 56780, 31200, None, 62100, 34500, None, None],
        ['Service Revenue - Training', 28900, 15600, None, 31200, 17800, None, None],
        ['Subscription Income - Basic', 45600, 12300, None, 48900, 13100, None, None],
        ['Subscription Income - Premium', 89200, 28700, None, 95600, 30200, None, None],
        ['Subscription Income - Enterprise', 156800, 52400, None, 168300, 55800, None, None],
        ['Licensing Fees - Software', 78300, 24500, None, 82100, 26200, None, None],
        ['Licensing Fees - Patents', 34200, 8900, None, 36800, 9400, None, None],
        ['Advertising Revenue', 23400, 14200, None, 26800, 15900, None, None],
        ['Commission Income', 18900, 7600, None, 21200, 8400, None, None],
        ['Rental Income - Equipment', 12400, 5800, None, 13600, 6200, None, None],
        ['Rental Income - Office Space', 67800, 42100, None, 67800, 43200, None, None],
        ['Government Contracts', 198500, 156200, None, 215600, 168900, None, None],
        ['International Sales - Europe', 124600, 78900, None, 138200, 85400, None, None],
        ['International Sales - Asia', 96300, 62100, None, 108400, 68900, None, None],
        ['International Sales - Americas', 78400, 48200, None, 85600, 52100, None, None],
        ['Wholesale Distribution', 234500, 189600, None, 248900, 198700, None, None],
        ['Retail Operations', 167800, 124500, None, 178200, 131600, None, None],
        ['E-commerce Sales', 89600, 45200, None, 102300, 49800, None, None],
        ['Partnership Revenue', 56700, 28400, None, 61200, 30600, None, None],
        ['Franchise Fees', 34800, 12100, None, 37200, 13400, None, None],
        ['Investment Returns', 28400, 5600, None, 31200, 6100, None, None],
        ['Insurance Proceeds', 15200, 3400, None, 8600, 2100, None, None],
        ['Miscellaneous Income', 8900, 4200, None, 9800, 4600, None, None],
        ['Interest Income', 12600, 2100, None, 13800, 2300, None, None],
        ['Dividend Income', 7800, 1500, None, 8400, 1700, None, None],
    ]

    currency_fmt = '#,##0'
    data_font = Font(name='Arial', size=10)
    data_align = Alignment(horizontal='right', vertical='center')
    cat_align = Alignment(horizontal='left', vertical='center')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 1:
                cell.alignment = cat_align
            else:
                cell.alignment = data_align
                if val is not None:
                    cell.number_format = currency_fmt

    # Set column widths
    ws.column_dimensions['A'].width = 35
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 15

    # Row height for header
    ws.row_dimensions[1].height = 30

    # Leave page margins at default (approximately 2cm / 0.75in top/bottom, 0.7in left/right)
    # The agent must change these to 1.5cm

    # Page setup: A4 Portrait
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
