"""
Reward Script: Use INDIRECT to reference a cell whose address is stored as text in another cell.
Task ID: calc_lf_025
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): E2 contains a formula (not empty, not plain value)
  Component 2 (0.3): The formula uses the INDIRECT function
  Component 3 (0.3): The INDIRECT formula references D2 (the cell containing the text address)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_025'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the Data sheet
    try:
        ws = wb['Data']
    except KeyError:
        # Try active sheet as fallback
        ws = wb.active
        if ws is None:
            print("CRITICAL: No 'Data' sheet and no active sheet found")
            print("REWARD: 0.0")
            return 0.0
        print(f"WARN: 'Data' sheet not found, using active sheet '{ws.title}'")

    # Component 1: E2 contains a formula (0.4 points)
    # In initial_env, E2 is None (empty). In golden_env, E2 should have a formula.
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str) and e2_value.startswith('='):
            print(f"PASS: Component 1 — E2 contains a formula: {e2_value} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — E2 expected a formula, found: {e2_value!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: The formula uses the INDIRECT function (0.3 points)
    # This specifically checks that INDIRECT is used, not just any formula.
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str):
            formula_upper = e2_value.upper().replace(' ', '')
            if 'INDIRECT(' in formula_upper:
                print(f"PASS: Component 2 — Formula uses INDIRECT function (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Formula does not use INDIRECT: {e2_value}")
        else:
            print(f"FAIL: Component 2 — E2 is not a formula string: {e2_value!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: INDIRECT references D2 specifically (0.3 points)
    # The task says D2 contains 'B4' and we should use INDIRECT(D2) to get the value in B4.
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str):
            formula_upper = e2_value.upper().replace(' ', '')
            # Accept =INDIRECT(D2) or =INDIRECT($D$2) or variations referencing D2
            if 'INDIRECT(' in formula_upper and 'D2' in formula_upper.replace('$', ''):
                print(f"PASS: Component 3 — INDIRECT references D2 (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — INDIRECT does not reference D2: {e2_value}")
        else:
            print(f"FAIL: Component 3 — E2 is not a formula string: {e2_value!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification (LibreOffice may have unsaved edits)
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
