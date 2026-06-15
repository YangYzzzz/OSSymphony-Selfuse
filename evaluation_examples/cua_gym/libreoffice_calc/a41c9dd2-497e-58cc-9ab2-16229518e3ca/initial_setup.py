"""
Initial Setup: Use INDIRECT to reference a cell whose address is stored as text.
Task ID: calc_lf_025
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Data'

    # Column A & B: Item / Value data
    ws['A1'] = 'Item'
    ws['B1'] = 'Value'
    ws['A2'] = 'X'
    ws['B2'] = 100
    ws['A3'] = 'Y'
    ws['B3'] = 200
    ws['A4'] = 'Z'
    ws['B4'] = 300

    # Column D & E: Cell Ref / Result
    ws['D1'] = 'Cell Ref'
    ws['D2'] = 'B4'
    ws['E1'] = 'Result'
    # E2 intentionally left empty — the agent must enter =INDIRECT(D2) here

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 4
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for the GUI agent
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
