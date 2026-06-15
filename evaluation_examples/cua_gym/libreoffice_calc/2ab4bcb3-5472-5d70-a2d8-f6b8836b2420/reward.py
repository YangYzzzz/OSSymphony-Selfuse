"""
Reward Script: Verify macro-applied bold + yellow background on cell B3
Task ID: calc_gg1_050
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Cell B3 has bold formatting
  Component 2 (0.5): Cell B3 has yellow background fill (solid, FFFFFF00)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_050'


def verify_task(file_path):
    """
    Verify that cell B3 has bold formatting and yellow background,
    confirming the macro was recorded and executed correctly.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: sheet has data (sanity check, not scored)
    if ws['A1'].value is None:
        print("CRITICAL: Sheet appears empty — precondition failed")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Cell B3 has bold formatting (0.5 points)
    # In initial_env, B3 is not bold. In golden_env, macro applied bold.
    try:
        cell_b3 = ws['B3']
        is_bold = cell_b3.font.bold is True
        if is_bold:
            print(f"PASS: Component 1 — B3 is bold (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — B3 bold={cell_b3.font.bold}, expected True")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cell B3 has yellow background fill (0.5 points)
    # In initial_env, B3 has no fill. In golden_env, macro applied yellow (FFFFFF00) solid fill.
    try:
        cell_b3 = ws['B3']
        fill_type = cell_b3.fill.fill_type
        fg_rgb = None
        try:
            fg_rgb = cell_b3.fill.fgColor.rgb
        except Exception:
            pass

        # Accept yellow variants: FFFFFF00 is pure yellow
        # Also accept common yellow shades (FFFFD700, FFFFFFE0, etc.) but primary target is FFFFFF00
        is_yellow_fill = False
        if fill_type == 'solid' and fg_rgb is not None:
            # Check for yellow: high R, high G, low B in ARGB format
            # FFFFFF00 = alpha(FF) R(FF) G(FF) B(00) — pure yellow
            rgb_str = str(fg_rgb).upper()
            if len(rgb_str) == 8:
                r_hex = rgb_str[2:4]
                g_hex = rgb_str[4:6]
                b_hex = rgb_str[6:8]
                try:
                    r_val = int(r_hex, 16)
                    g_val = int(g_hex, 16)
                    b_val = int(b_hex, 16)
                    # Yellow: R >= 200, G >= 200, B <= 100
                    if r_val >= 200 and g_val >= 200 and b_val <= 100:
                        is_yellow_fill = True
                except ValueError:
                    pass

        if is_yellow_fill:
            print(f"PASS: Component 2 — B3 has yellow background fill (fill_type={fill_type}, fgColor={fg_rgb}) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — B3 fill_type={fill_type}, fgColor={fg_rgb}, expected solid yellow")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
