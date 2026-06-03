"""
Initial Setup: Record a macro for bold + yellow formatting, run on test cell
Task ID: calc_gg1_050
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: DataReview ---
    ws = wb.active
    ws.title = 'DataReview'

    # Headers
    headers = ['Employee', 'Score', 'Department', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 16

    # Realistic employee review data (10 rows)
    data = [
        ['Sarah Chen',        87, 'Engineering',   'Reviewed'],
        ['Marcus Johnson',    92, 'Marketing',     'Pending'],
        ['Elena Rodriguez',   78, 'Finance',       'Reviewed'],
        ['James O\'Brien',    95, 'Engineering',   'Approved'],
        ['Priya Patel',       83, 'Human Resources','Pending'],
        ['David Kim',         71, 'Operations',    'Reviewed'],
        ['Lisa Nakamura',     88, 'Marketing',     'Approved'],
        ['Robert Okafor',     76, 'Finance',       'Pending'],
        ['Maria Gonzalez',    90, 'Engineering',   'Reviewed'],
        ['Thomas Wright',     84, 'Operations',    'Approved'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # No macros, no special formatting on B3 - just plain data
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
