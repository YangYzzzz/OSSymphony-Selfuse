"""
Reward Script: Timesheet data cleanup — normalize hour formats, add SUM totals, flag >50h in orange
Task ID: calc_gen_data_cleanup_013
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): B2:F26 all converted to numeric decimals (no text suffixes)
  - Component 2 (0.3): G2:G26 contain =SUM(B_:F_) formulas
  - Component 3 (0.2): Conditional formatting on G2:G26 — orange fill for values > 50
  - Component 4 (0.1): Column G number format is '0.0' (1 decimal place)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_013'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Timesheet' not in wb.sheetnames:
        print("CRITICAL: 'Timesheet' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timesheet']

    # Component 1: B2:F26 all converted to numeric decimals, no text strings (0.4 points)
    # The initial file has ~123 text-formatted cells like '8h', '8 hours', '8:00'
    # The golden file converts all to plain numeric values (int or float)
    try:
        text_cells = []
        non_numeric_count = 0
        total_cells = 0
        for row_idx in range(2, 27):
            for col_idx in range(2, 7):  # columns B through F
                cell = ws.cell(row=row_idx, column=col_idx)
                v = cell.value
                total_cells += 1
                if isinstance(v, str):
                    # Any string value indicates text format (not cleaned up)
                    non_numeric_count += 1
                    if len(text_cells) < 5:
                        from openpyxl.utils import get_column_letter
                        text_cells.append(f"{get_column_letter(col_idx)}{row_idx}={repr(v)}")
                elif v is None:
                    # None is also problematic — missing value
                    non_numeric_count += 1
                # int or float values are correctly cleaned

        if non_numeric_count == 0:
            print(f"PASS: Component 1 — all {total_cells} cells in B2:F26 are numeric (0.4 pts)")
            total_score += 0.4
        else:
            examples = ', '.join(text_cells) if text_cells else 'none sampled'
            print(f"FAIL: Component 1 — {non_numeric_count}/{total_cells} cells in B2:F26 still have"
                  f" non-numeric values. Examples: {examples}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: G2:G26 contain =SUM(Bx:Fx) formulas (0.3 points)
    # The initial file has all None in column G; the golden file has SUM formulas for each row
    try:
        sum_formula_count = 0
        wrong_formulas = []
        for row_idx in range(2, 27):
            v = ws.cell(row=row_idx, column=7).value  # column G
            if isinstance(v, str):
                v_clean = v.upper().replace(' ', '')
                expected = f'=SUM(B{row_idx}:F{row_idx})'
                expected_clean = expected.upper().replace(' ', '')
                if v_clean == expected_clean:
                    sum_formula_count += 1
                else:
                    wrong_formulas.append(f"G{row_idx}={repr(v)} (expected {expected})")
            else:
                wrong_formulas.append(f"G{row_idx}={repr(v)} (expected =SUM formula)")

        if sum_formula_count == 25:
            print(f"PASS: Component 2 — all 25 rows in G2:G26 have =SUM(Bx:Fx) formulas (0.3 pts)")
            total_score += 0.3
        elif sum_formula_count >= 20:
            # Partial: most formulas correct
            print(f"PARTIAL: Component 2 — {sum_formula_count}/25 rows have correct SUM formulas."
                  f" Wrong: {wrong_formulas[:3]}")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — only {sum_formula_count}/25 rows have =SUM formulas."
                  f" First wrong: {wrong_formulas[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on G2:G26 with orange fill when value > 50 (0.2 points)
    # The initial file has no conditional formatting; the golden has cellIs > 50 with orange fill
    try:
        cf_found = False
        cf_range_correct = False
        cf_operator_correct = False
        cf_threshold_correct = False
        cf_fill_orange = False

        for cf in ws.conditional_formatting:
            sqref_str = str(cf.sqref)
            # Check it covers G2:G26
            if 'G2' in sqref_str and 'G26' in sqref_str:
                cf_range_correct = True

            for rule in cf.rules:
                if rule.type == 'cellIs':
                    cf_found = True
                    # Check operator is greaterThan
                    if getattr(rule, 'operator', None) == 'greaterThan':
                        cf_operator_correct = True
                    # Check threshold is 50
                    formula = getattr(rule, 'formula', [])
                    if formula and str(formula[0]) == '50':
                        cf_threshold_correct = True
                    # Check fill is orange
                    if hasattr(rule, 'dxf') and rule.dxf is not None:
                        dxf = rule.dxf
                        if hasattr(dxf, 'fill') and dxf.fill is not None:
                            try:
                                fg_rgb = dxf.fill.fgColor.rgb
                                # Orange fill: FFFF6600 (opaque orange)
                                if fg_rgb and fg_rgb.upper() in ('FFFF6600', 'FF6600', 'FFA500',
                                                                  'FFFFA500', 'FFFF9900'):
                                    cf_fill_orange = True
                                elif fg_rgb:
                                    # Accept any reddish-orange color range
                                    # Check R is high (FF), G moderate (40-A0), B is low (00-30)
                                    try:
                                        argb = fg_rgb.upper().lstrip('0')
                                        if len(fg_rgb) == 8:
                                            r = int(fg_rgb[2:4], 16)
                                            g = int(fg_rgb[4:6], 16)
                                            b = int(fg_rgb[6:8], 16)
                                            if r >= 200 and g >= 80 and b <= 50:
                                                cf_fill_orange = True
                                    except:
                                        pass
                                if not cf_fill_orange:
                                    print(f"  CF fill color found: {fg_rgb} (expected FFFF6600 orange)")
                            except Exception as ce:
                                print(f"  CF fill check error: {ce}")

        if cf_found and cf_range_correct and cf_operator_correct and cf_threshold_correct and cf_fill_orange:
            print(f"PASS: Component 3 — conditional formatting on G2:G26, cellIs > 50, orange fill (0.2 pts)")
            total_score += 0.2
        elif cf_found and cf_range_correct and cf_operator_correct and cf_threshold_correct:
            print(f"PARTIAL: Component 3 — CF rule found (range={cf_range_correct},"
                  f" op={cf_operator_correct}, threshold={cf_threshold_correct})"
                  f" but orange fill not confirmed")
            total_score += 0.1
        else:
            print(f"FAIL: Component 3 — CF check: found={cf_found}, range={cf_range_correct},"
                  f" operator={cf_operator_correct}, threshold={cf_threshold_correct},"
                  f" orange={cf_fill_orange}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Column G number format is '0.0' (1 decimal place) (0.1 points)
    # The initial file has no number format on G (all None); the golden has '0.0' format
    try:
        fmt_correct_count = 0
        fmt_sample = []
        for row_idx in range(2, 27):
            cell = ws.cell(row=row_idx, column=7)
            nf = cell.number_format
            # Accept '0.0' or similar 1-decimal-place formats
            if nf and nf.strip() in ('0.0', '#,##0.0', '0.0_)', '0.00'):
                fmt_correct_count += 1
            elif nf == '0.0':
                fmt_correct_count += 1
            else:
                if len(fmt_sample) < 3:
                    fmt_sample.append(f"G{row_idx}='{nf}'")

        if fmt_correct_count == 25:
            print(f"PASS: Component 4 — all 25 G-column cells have '0.0' number format (0.1 pts)")
            total_score += 0.1
        elif fmt_correct_count >= 20:
            print(f"PARTIAL: Component 4 — {fmt_correct_count}/25 cells have correct number format")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — only {fmt_correct_count}/25 G cells have '0.0' format."
                  f" Sample wrong: {fmt_sample}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
