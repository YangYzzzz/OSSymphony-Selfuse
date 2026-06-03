"""
Initial Setup: Timesheet with mixed-format hour entries
Task ID: calc_gen_data_cleanup_013
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timesheet'

    # --- Row 1: Headers ---
    headers = ['Employee', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Total Hours']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 25 employees with mixed-format hours in B-F
    # Formats used: '8h', '8 hours', '8.0', '8:00', '7.5h', plain float/int
    # Column G is left empty (no formulas, no totals)
    employees_data = [
        # (Name, Mon, Tue, Wed, Thu, Fri)  — all B-F as mixed strings/numbers
        ('Sarah Chen',       '8h',       '7.5h',      '8.0',       '8:00',      '6h'),
        ('Marcus Johnson',   '8 hours',  8.0,         '7:30',      '8h',        '7.5h'),
        ('Emily Rodriguez',  '9h',       '9 hours',   '8.5',       '8:00',      '9h'),
        ('David Kim',        '7:00',     '8h',        '8 hours',   '7.5',       '8.0'),
        ('Rachel Thompson',  8.0,        '8:00',      '8h',        '7 hours',   '8h'),
        ('James Patel',      '10h',      '10 hours',  '10.0',      '10:00',     '10h'),
        ('Lauren White',     '6h',       '5.5h',      '6.0',       '6:00',      '5h'),
        ('Michael Torres',   '8:30',     '8h',        '9 hours',   '8.5',       '8.0'),
        ('Jessica Brown',    '7.5h',     '7:30',      '8h',        '8 hours',   '7.5'),
        ('Kevin Nguyen',     '9h',       '8.5',       '9:00',      '8h',        '9 hours'),
        ('Amanda Foster',    '11h',      '11 hours',  '11.0',      '11:00',     '10h'),
        ('Brian Mitchell',   '8h',       '8.0',       '8 hours',   '8:00',      '7h'),
        ('Stephanie Clark',  '7:00',     '7h',        '7.5',       '7 hours',   '8.0'),
        ('Christopher Lee',  '9:30',     '9h',        '9 hours',   '9.5',       '8h'),
        ('Nicole Harris',    '6h',       '6.0',       '6:30',      '7h',        '6 hours'),
        ('Robert Wilson',    '10h',      '10.0',      '10 hours',  '9:30',      '10h'),
        ('Melissa Adams',    '8:00',     '8h',        '8.5',       '8 hours',   '8.0'),
        ('Daniel Martinez',  '7h',       '7.5h',      '7:30',      '7.0',       '7 hours'),
        ('Ashley Jackson',   '9h',       '9.0',       '9:00',      '9 hours',   '8.5h'),
        ('Tyler Robinson',   '11h',      '10.5',      '11 hours',  '10:30',     '10h'),
        ('Samantha Davis',   '8h',       '8:00',      '8.0',       '8 hours',   '7.5'),
        ('Nathan Moore',     '6.5h',     '6:30',      '6h',        '6 hours',   '6.5'),
        ('Brittany Taylor',  '9:00',     '9h',        '8.5',       '9 hours',   '9.0'),
        ('Eric Anderson',    '10h',      '10 hours',  '10:00',     '10.0',      '9h'),
        ('Heather Thomas',   '7.5h',     '7h',        '7 hours',   '7:30',      '8.0'),
    ]

    for r, (name, mon, tue, wed, thu, fri) in enumerate(employees_data, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=mon)
        ws.cell(row=r, column=3, value=tue)
        ws.cell(row=r, column=4, value=wed)
        ws.cell(row=r, column=5, value=thu)
        ws.cell(row=r, column=6, value=fri)
        # Column G (Total Hours) intentionally left empty

    # Set some column widths for readability
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
