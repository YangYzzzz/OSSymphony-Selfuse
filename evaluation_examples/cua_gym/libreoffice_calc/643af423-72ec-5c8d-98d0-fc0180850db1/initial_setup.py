"""
Initial Setup: Student demographics database for equity report filter task
Task ID: calc_edu_student_demographics_filter_029
Domain: libreoffice_calc
"""

import random
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_student_demographics_filter_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

STEM_MAJORS = ['Computer Science', 'Biology', 'Mathematics', 'Physics', 'Engineering']
NON_STEM_MAJORS = ['History', 'English', 'Business', 'Political Science', 'Psychology',
                   'Sociology', 'Art', 'Music', 'Philosophy', 'Economics']
ALL_MAJORS = STEM_MAJORS + NON_STEM_MAJORS

FIRST_NAMES = [
    'Emma', 'Liam', 'Olivia', 'Noah', 'Ava', 'Elijah', 'Isabella', 'James',
    'Sophia', 'Oliver', 'Mia', 'Benjamin', 'Charlotte', 'Lucas', 'Amelia',
    'Mason', 'Harper', 'Ethan', 'Evelyn', 'Alexander', 'Abigail', 'Henry',
    'Emily', 'Sebastian', 'Elizabeth', 'Michael', 'Mila', 'Daniel', 'Ella',
    'Owen', 'Chloe', 'Logan', 'Victoria', 'Jackson', 'Sofia', 'Aiden',
    'Grace', 'Carter', 'Lily', 'Jayden', 'Penelope', 'Gabriel', 'Layla',
    'Ryan', 'Nora', 'Dylan', 'Zoey', 'Nathan', 'Riley', 'Caleb',
    'Hannah', 'Leo', 'Lillian', 'Isaiah', 'Addison', 'Joshua', 'Aurora',
    'Andrew', 'Savannah', 'Lincoln', 'Brooklyn', 'Anthony', 'Bella',
    'Kai', 'Zoe', 'Aaron', 'Maya', 'David', 'Luna', 'Tyler', 'Skylar',
    'Connor', 'Naomi', 'Evan', 'Aaliyah', 'Jordan', 'Elena', 'Kevin', 'Anna',
    'Miguel', 'Sofia', 'Carlos', 'Valentina', 'Mateo', 'Isabella',
    'Wei', 'Mei', 'Jun', 'Lin', 'Raj', 'Priya', 'Arjun', 'Aisha',
    'Zara', 'Kenji', 'Yuki', 'Sana', 'Fatima', 'Omar', 'Leila', 'Hassan'
]

LAST_NAMES = [
    'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
    'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
    'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
    'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
    'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
    'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green',
    'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
    'Carter', 'Roberts', 'Chen', 'Zhang', 'Wang', 'Kim', 'Patel',
    'Singh', 'Kumar', 'Sharma', 'Ahmed', 'Ali', 'Khan', 'Hassan',
    'Yamamoto', 'Tanaka', 'Nakamura', 'Suzuki', 'Watanabe', 'Ito',
    'Okafor', 'Mensah', 'Diallo', 'Tremblay', 'Leblanc', 'Dubois'
]

STATUSES = ['Enrolled', 'Enrolled', 'Enrolled', 'Enrolled', 'Part-Time', 'Part-Time', 'Leave of Absence']

PELL_BY_FIRST_GEN = {
    'Yes': ['Yes', 'Yes', 'Yes', 'No'],   # First-gen more likely to have Pell
    'No': ['Yes', 'No', 'No', 'No']
}


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Demographics'

    # --- Headers ---
    headers = ['Student ID', 'Name', 'Major', 'First Gen', 'Pell Grant', 'GPA', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Generate 300 student records ---
    used_ids = set()
    used_names = set()

    for i in range(300):
        row = i + 2

        # Unique student ID
        while True:
            sid = f'STU{random.randint(10000, 99999)}'
            if sid not in used_ids:
                used_ids.add(sid)
                break

        # Unique name
        while True:
            fn = random.choice(FIRST_NAMES)
            ln = random.choice(LAST_NAMES)
            full = f'{fn} {ln}'
            if full not in used_names:
                used_names.add(full)
                break

        # Major — distribute roughly: ~40% STEM, ~60% non-STEM
        if random.random() < 0.40:
            major = random.choice(STEM_MAJORS)
        else:
            major = random.choice(NON_STEM_MAJORS)

        # First Gen — roughly 35% first-gen
        first_gen = 'Yes' if random.random() < 0.35 else 'No'

        # Pell Grant
        pell = random.choice(PELL_BY_FIRST_GEN[first_gen])

        # GPA — 2.0 to 4.0 with realistic distribution
        gpa = round(random.uniform(2.0, 4.0), 2)

        # Status
        status = random.choice(STATUSES)

        ws.cell(row=row, column=1, value=sid)
        ws.cell(row=row, column=2, value=full)
        ws.cell(row=row, column=3, value=major)
        ws.cell(row=row, column=4, value=first_gen)
        ws.cell(row=row, column=5, value=pell)
        ws.cell(row=row, column=6, value=gpa)
        ws.cell(row=row, column=7, value=status)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Demographics, 300 student rows (rows 2-301)')
    print(f'  Columns: Student ID, Name, Major, First Gen, Pell Grant, GPA, Status')
    print(f'  No AutoFilter set (task requires agent to apply it)')
    print(f'  No STEM First-Gen sheet (task requires agent to create it)')


create_initial()
