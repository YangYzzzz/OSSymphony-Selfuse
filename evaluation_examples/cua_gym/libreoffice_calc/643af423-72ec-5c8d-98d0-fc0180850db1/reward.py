"""
Reward Script: Apply filters for first-generation STEM students and copy visible rows to new sheet
Task ID: calc_edu_student_demographics_filter_029
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: 'STEM First-Gen' sheet exists in the workbook             — 0.30 pts
  Component 2: 'STEM First-Gen' header row matches Demographics headers  — 0.20 pts
  Component 3: All data rows have STEM major AND First Gen = 'Yes'       — 0.30 pts
  Component 4: Correct number of qualifying rows (44 data rows)          — 0.20 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_student_demographics_filter_029'

STEM_MAJORS = {'Computer Science', 'Biology', 'Mathematics', 'Physics', 'Engineering'}
EXPECTED_HEADERS = ['Student ID', 'Name', 'Major', 'First Gen', 'Pell Grant', 'GPA', 'Status']
# Expected count of first-gen STEM students in the dataset (determined from golden file analysis)
EXPECTED_DATA_ROWS = 44


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: 'STEM First-Gen' sheet exists (0.30 points)
    # This FAILS on initial (no such sheet) and PASSES on golden (sheet present)
    # -----------------------------------------------------------------------
    try:
        if 'STEM First-Gen' in wb.sheetnames:
            print("PASS: Component 1 — 'STEM First-Gen' sheet exists (0.30 pts)")
            total_score += 0.30
            ws_stem = wb['STEM First-Gen']
        else:
            print(f"FAIL: Component 1 — 'STEM First-Gen' sheet not found. Sheets present: {wb.sheetnames}")
            # Cannot proceed with further checks
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # -----------------------------------------------------------------------
    # Component 2: Header row in 'STEM First-Gen' matches Demographics headers (0.20 points)
    # The new sheet must include the same 7-column header row as the source data.
    # -----------------------------------------------------------------------
    try:
        actual_headers = []
        for col in range(1, 8):
            val = ws_stem.cell(row=1, column=col).value
            actual_headers.append(val)

        if actual_headers == EXPECTED_HEADERS:
            print(f"PASS: Component 2 — Header row matches expected: {actual_headers} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Header mismatch.")
            print(f"  Expected: {EXPECTED_HEADERS}")
            print(f"  Actual:   {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: All data rows in 'STEM First-Gen' have STEM major AND First Gen = 'Yes' (0.30 points)
    # This is the core task requirement — only qualifying students should be in this sheet.
    # -----------------------------------------------------------------------
    try:
        max_row = ws_stem.max_row
        if max_row < 2:
            print("FAIL: Component 3 — No data rows in 'STEM First-Gen' sheet")
        else:
            invalid_count = 0
            invalid_examples = []
            for row in range(2, max_row + 1):
                major = ws_stem.cell(row=row, column=3).value
                first_gen = ws_stem.cell(row=row, column=4).value
                if major not in STEM_MAJORS or first_gen != 'Yes':
                    invalid_count += 1
                    if len(invalid_examples) < 5:
                        invalid_examples.append({
                            'row': row,
                            'major': major,
                            'first_gen': first_gen
                        })

            if invalid_count == 0:
                print(f"PASS: Component 3 — All {max_row - 1} data rows have STEM major AND First Gen = 'Yes' (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 3 — Found {invalid_count} invalid rows (showing up to 5):")
                for r in invalid_examples:
                    print(f"  Row {r['row']}: Major={r['major']}, First Gen={r['first_gen']}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Correct number of qualifying rows — 44 data rows (0.20 points)
    # Based on the 300 student records, exactly 44 are first-gen STEM students.
    # -----------------------------------------------------------------------
    try:
        actual_data_rows = ws_stem.max_row - 1  # Subtract header row
        if actual_data_rows == EXPECTED_DATA_ROWS:
            print(f"PASS: Component 4 — Correct row count: {actual_data_rows} data rows (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Expected {EXPECTED_DATA_ROWS} data rows, found {actual_data_rows}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
