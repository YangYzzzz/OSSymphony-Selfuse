"""
Initial Setup: WEF Annual Davos Forum tracker — Location column blank
Task ID: osworld_multi_apps_conference_city_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_multi_apps_conference_city_007'
OUTPUT = f'{WORKDIR}/WEF.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: WEF Annual Davos Forum ---
    ws = wb.active
    ws.title = 'WEF'

    # Header row
    ws.cell(row=1, column=1, value='Year')
    ws.cell(row=1, column=2, value='Location')

    # Style headers
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows: years 2015-2023, Location intentionally BLANK
    years = [2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022, 2023]
    for i, year in enumerate(years, start=2):
        ws.cell(row=i, column=1, value=year)
        # Column B (Location) is left blank — task is to fill it in

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25

    # Row height for header
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open WEF.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with WEF.xlsx on DISPLAY=:0')


create_initial()
