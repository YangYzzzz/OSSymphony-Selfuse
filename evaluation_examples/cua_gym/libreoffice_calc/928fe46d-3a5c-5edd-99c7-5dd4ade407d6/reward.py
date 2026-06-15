"""
Reward Script: Fill WEF host locations for 2015-2023
Task ID: osworld_multi_apps_conference_city_007
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4 pts): "Davos, Switzerland" for all years except 2021 (rows 2-7, 9-10)
  - Component 2 (0.3 pts): "Virtual" for 2021 (COVID year)
  - Component 3 (0.3 pts): All 9 location cells are non-blank (full completion gate)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_007'

# Expected location data: year -> location
EXPECTED_LOCATIONS = {
    2015: 'Davos, Switzerland',
    2016: 'Davos, Switzerland',
    2017: 'Davos, Switzerland',
    2018: 'Davos, Switzerland',
    2019: 'Davos, Switzerland',
    2020: 'Davos, Switzerland',
    2021: 'Virtual',
    2022: 'Davos, Switzerland',
    2023: 'Davos, Switzerland',
}

DAVOS_YEARS = [2015, 2016, 2017, 2018, 2019, 2020, 2022, 2023]
VIRTUAL_YEAR = 2021


def normalize(val):
    """Normalize a cell value for comparison: strip whitespace, lowercase."""
    if val is None:
        return ''
    return str(val).strip()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Access the WEF sheet
    try:
        if 'WEF' in wb.sheetnames:
            ws = wb['WEF']
        else:
            ws = wb.worksheets[0]
            print(f"WARN: Sheet 'WEF' not found, using sheet '{ws.title}'")
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Build a mapping of year -> location from the file
    # Expected: row 1 = headers (Year, Location), rows 2-10 = data
    year_to_location = {}
    try:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
            year_cell = row[0]
            loc_cell = row[1]
            if year_cell.value is not None:
                try:
                    year = int(year_cell.value)
                    loc = normalize(loc_cell.value)
                    year_to_location[year] = loc
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        print(f"CRITICAL: Cannot read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    print(f"Parsed year-location mapping: {year_to_location}")

    # --- Component 1: 'Davos, Switzerland' for all non-2021 years (0.4 points) ---
    # These entries should fail on initial (all blank) and pass on golden
    try:
        davos_correct = 0
        davos_total = len(DAVOS_YEARS)
        for year in DAVOS_YEARS:
            loc = year_to_location.get(year, '')
            if loc.lower() == 'davos, switzerland':
                davos_correct += 1
            else:
                print(f"FAIL Comp1: Year {year} expected 'Davos, Switzerland', found '{loc}'")

        if davos_correct == davos_total:
            print(f"PASS: Component 1 — All {davos_total} Davos years correctly set to 'Davos, Switzerland' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Only {davos_correct}/{davos_total} Davos years correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: 'Virtual' for 2021 (0.3 points) ---
    # The COVID exception year — held online
    try:
        loc_2021 = year_to_location.get(2021, '')
        if loc_2021.lower() == 'virtual':
            print(f"PASS: Component 2 — 2021 correctly set to 'Virtual' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — 2021 expected 'Virtual', found '{loc_2021}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: All 9 location cells are non-blank (0.3 points) ---
    # Verifies that ALL years have been filled in, not just some
    try:
        all_years = list(range(2015, 2024))  # 2015-2023 inclusive
        filled = sum(1 for y in all_years if year_to_location.get(y, ''))
        if filled == 9:
            print(f"PASS: Component 3 — All 9 years have non-blank locations (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — Only {filled}/9 years have non-blank locations")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/WEF.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
