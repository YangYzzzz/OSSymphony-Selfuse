"""
Reward Script: KPI Dashboard with IF formulas and conditional formatting
Task ID: calc_ops_037
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): IF formulas present in D2:D6 with correct logic
  Component 2 (0.3): Formula results match ground truth values
  Component 3 (0.3): Conditional formatting rules on D2:D6 (green=Met, red=Below Target)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_037'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (formula mode)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: KPI sheet must exist
    if 'KPI' not in wb.sheetnames:
        print("FAIL: 'KPI' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['KPI']

    # =========================================================
    # Component 1: IF formulas in D2:D6 with correct logic (0.4 pts)
    # D2-D4 should use >= (higher is better), D5-D6 should use <= (lower is better)
    # =========================================================
    try:
        formulas_correct = 0
        expected_patterns = {
            2: ('>=', 'higher-is-better: On-Time Delivery'),
            3: ('>=', 'higher-is-better: Fill Rate'),
            4: ('>=', 'higher-is-better: Inventory Accuracy'),
            5: ('<=', 'lower-is-better: Order Cycle Time'),
            6: ('<=', 'lower-is-better: Defect Rate'),
        }

        for row, (expected_op, desc) in expected_patterns.items():
            cell_val = ws.cell(row=row, column=4).value
            if cell_val is None:
                print(f"FAIL: D{row} is empty (expected IF formula for {desc})")
                continue

            val_str = str(cell_val).upper().replace(' ', '')
            # Must be an IF formula
            if not val_str.startswith('=IF('):
                print(f"FAIL: D{row} is not an IF formula: {cell_val}")
                continue

            # Check correct comparison operator
            # For >=: formula should contain C{row}>=B{row} or equivalent
            # For <=: formula should contain C{row}<=B{row} or equivalent
            has_correct_op = False
            if expected_op == '>=':
                # Accept C>=B patterns (actual >= target means Met)
                if f'C{row}>=B{row}' in val_str or f'C{row}>B{row}' in val_str:
                    has_correct_op = True
                # Also accept B<=C patterns
                if f'B{row}<=C{row}' in val_str or f'B{row}<C{row}' in val_str:
                    has_correct_op = True
            elif expected_op == '<=':
                # Accept C<=B patterns (actual <= target means Met for lower-is-better)
                if f'C{row}<=B{row}' in val_str or f'C{row}<B{row}' in val_str:
                    has_correct_op = True
                # Also accept B>=C patterns
                if f'B{row}>=C{row}' in val_str or f'B{row}>C{row}' in val_str:
                    has_correct_op = True

            # Must contain "Met" and "Below Target" (or similar)
            has_met = '"MET"' in val_str
            has_below = '"BELOWTARGET"' in val_str

            if has_correct_op and has_met and has_below:
                formulas_correct += 1
                print(f"PASS: D{row} has correct IF formula ({desc}): {cell_val}")
            elif has_met and has_below:
                # Has IF with Met/Below Target but wrong operator — partial
                formulas_correct += 0.5
                print(f"PARTIAL: D{row} has IF formula with Met/Below Target but operator may be wrong: {cell_val}")
            else:
                print(f"FAIL: D{row} formula incomplete: {cell_val}")

        comp1_score = (formulas_correct / 5.0) * 0.4
        total_score += comp1_score
        print(f"Component 1 subtotal: {comp1_score:.2f}/0.40 ({formulas_correct}/5 formulas correct)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================
    # Component 2: Formulas would produce correct results (0.3 pts)
    # Ground truth: D2=Met, D3=Below Target, D4=Met, D5=Below Target, D6=Met
    # Since openpyxl cannot evaluate formulas and data_only returns None for
    # files not yet opened in LibreOffice, we verify that the formula logic
    # combined with actual B/C values would produce the correct result.
    # For rows 2-4 (higher is better): Met if C>=B
    # For rows 5-6 (lower is better): Met if C<=B
    # =========================================================
    try:
        expected_results = {
            2: ('Met', '>='),       # 0.97 >= 0.95 -> Met
            3: ('Below Target', '>='),  # 0.96 >= 0.98 -> Below Target (fails)
            4: ('Met', '>='),       # 0.995 >= 0.99 -> Met
            5: ('Below Target', '<='),  # 3.5 <= 3 -> Below Target (fails)
            6: ('Met', '<='),       # 0.015 <= 0.02 -> Met
        }

        results_correct = 0
        for row, (expected_val, op) in expected_results.items():
            formula = ws.cell(row=row, column=4).value
            target = ws.cell(row=row, column=2).value
            actual = ws.cell(row=row, column=3).value

            if formula is None or target is None or actual is None:
                print(f"FAIL: D{row} missing formula or B/C values")
                continue

            formula_str = str(formula).upper().replace(' ', '')
            if not formula_str.startswith('=IF('):
                print(f"FAIL: D{row} not an IF formula, cannot verify result")
                continue

            # Evaluate what the formula would produce based on B and C values
            try:
                t = float(target)
                a = float(actual)
            except (ValueError, TypeError):
                print(f"FAIL: D{row} B/C values are not numeric: B={target}, C={actual}")
                continue

            # Determine if formula uses correct operator for this row
            if op == '>=':
                # For higher-is-better: Met if actual >= target
                would_produce_met = (a >= t)
                # Check formula has >= or > for C vs B
                correct_direction = (f'C{row}>=B{row}' in formula_str or
                                     f'C{row}>B{row}' in formula_str or
                                     f'B{row}<=C{row}' in formula_str or
                                     f'B{row}<C{row}' in formula_str)
            else:
                # For lower-is-better: Met if actual <= target
                would_produce_met = (a <= t)
                # Check formula has <= or < for C vs B
                correct_direction = (f'C{row}<=B{row}' in formula_str or
                                     f'C{row}<B{row}' in formula_str or
                                     f'B{row}>=C{row}' in formula_str or
                                     f'B{row}>C{row}' in formula_str)

            predicted_result = 'Met' if (correct_direction and would_produce_met) or (not correct_direction and not would_produce_met) else 'Below Target'
            # Actually, let's be more precise: if the formula direction is correct,
            # the predicted result follows the data. If not, it's inverted.
            # Simpler: check if the formula would yield the expected result
            if correct_direction:
                predicted_result = 'Met' if would_produce_met else 'Below Target'
            else:
                # Wrong direction in formula means inverted result
                predicted_result = 'Below Target' if would_produce_met else 'Met'

            if predicted_result == expected_val:
                results_correct += 1
                print(f"PASS: D{row} formula would produce '{expected_val}' (target={t}, actual={a})")
            else:
                print(f"FAIL: D{row} formula would produce '{predicted_result}', expected '{expected_val}'")

        comp2_score = (results_correct / 5.0) * 0.3
        total_score += comp2_score
        print(f"Component 2 subtotal: {comp2_score:.2f}/0.30 ({results_correct}/5 results correct)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================
    # Component 3: Conditional formatting rules on D2:D6 (0.3 pts)
    # Green fill for "Met", red fill for "Below Target"
    # =========================================================
    try:
        has_green_rule = False
        has_red_rule = False
        cf_covers_d_range = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the CF range covers D2:D6 area
            if 'D' in cf_range:
                cf_covers_d_range = True
                for rule in cf.rules:
                    # Check for "Met" rule with green fill
                    if rule.formula and any('"Met"' in f for f in rule.formula):
                        if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                            rgb = rule.dxf.fill.fgColor.rgb
                            # Accept various greens
                            if rgb and ('00B050' in str(rgb) or '00FF00' in str(rgb) or '92D050' in str(rgb)):
                                has_green_rule = True
                                print(f"PASS: Green conditional formatting for 'Met' found (color: {rgb})")
                            else:
                                print(f"PARTIAL: 'Met' rule found but fill color is {rgb}, not green")
                        else:
                            print("PARTIAL: 'Met' rule found but no fill defined")

                    # Check for "Below Target" rule with red fill
                    if rule.formula and any('"Below Target"' in f for f in rule.formula):
                        if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                            rgb = rule.dxf.fill.fgColor.rgb
                            # Accept various reds
                            if rgb and ('FF0000' in str(rgb) or 'FF4444' in str(rgb) or 'C00000' in str(rgb)):
                                has_red_rule = True
                                print(f"PASS: Red conditional formatting for 'Below Target' found (color: {rgb})")
                            else:
                                print(f"PARTIAL: 'Below Target' rule found but fill color is {rgb}, not red")
                        else:
                            print("PARTIAL: 'Below Target' rule found but no fill defined")

                    # Also check cellIs rules with equal operator
                    if hasattr(rule, 'operator') and rule.operator == 'equal':
                        if rule.formula:
                            formula_str = str(rule.formula)
                            if '"Met"' in formula_str:
                                if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                                    rgb = rule.dxf.fill.fgColor.rgb
                                    if rgb and ('00B050' in str(rgb) or '00FF00' in str(rgb) or '92D050' in str(rgb)):
                                        has_green_rule = True
                                        print(f"PASS: Green cellIs rule for 'Met' found (color: {rgb})")
                            if '"Below Target"' in formula_str:
                                if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                                    rgb = rule.dxf.fill.fgColor.rgb
                                    if rgb and ('FF0000' in str(rgb) or 'FF4444' in str(rgb) or 'C00000' in str(rgb)):
                                        has_red_rule = True
                                        print(f"PASS: Red cellIs rule for 'Below Target' found (color: {rgb})")

        comp3_score = 0.0
        if cf_covers_d_range:
            if has_green_rule:
                comp3_score += 0.15
            else:
                print("FAIL: No green conditional formatting rule for 'Met'")
            if has_red_rule:
                comp3_score += 0.15
            else:
                print("FAIL: No red conditional formatting rule for 'Below Target'")
        else:
            print("FAIL: No conditional formatting found covering D column")

        total_score += comp3_score
        print(f"Component 3 subtotal: {comp3_score:.2f}/0.30")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
