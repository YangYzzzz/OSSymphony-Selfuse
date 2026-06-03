"""
Initial Setup: Build a KPI dashboard summary with conditional formatting
Task ID: calc_ops_037
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_037'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KPI'

    # --- Headers ---
    headers = ['KPI', 'Target', 'Actual', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- KPI Data (no formulas in D column - that's the task) ---
    data = [
        ['On-Time Delivery', 0.95, 0.97],
        ['Fill Rate', 0.98, 0.96],
        ['Inventory Accuracy', 0.99, 0.995],
        ['Order Cycle Time (days)', 3, 3.5],
        ['Defect Rate', 0.02, 0.015],
    ]

    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    pct_format = '0.00%'
    num_format = '0.0'

    for r, row_data in enumerate(data, 2):
        # Column A: KPI name
        cell_a = ws.cell(row=r, column=1, value=row_data[0])
        cell_a.font = Font(name='Calibri', size=11)
        cell_a.border = data_border

        # Column B: Target
        cell_b = ws.cell(row=r, column=2, value=row_data[1])
        cell_b.border = data_border
        cell_b.alignment = Alignment(horizontal='center')

        # Column C: Actual
        cell_c = ws.cell(row=r, column=3, value=row_data[2])
        cell_c.border = data_border
        cell_c.alignment = Alignment(horizontal='center')

        # Format percentages vs plain numbers
        if row_data[0] in ('Order Cycle Time (days)',):
            cell_b.number_format = num_format
            cell_c.number_format = num_format
        else:
            cell_b.number_format = pct_format
            cell_c.number_format = pct_format

        # Column D: Status - leave EMPTY (task is to add formulas here)
        cell_d = ws.cell(row=r, column=4)
        cell_d.border = data_border
        cell_d.alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
