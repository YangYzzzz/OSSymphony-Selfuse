"""
Reward Script: Goal Seek Break-Even Analysis
Task ID: calc_gg3_027
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): C3 on 'Break-Even' sheet contains the break-even value 1250
  Component 2 (0.4): Net income computes to 0 given the values in C2-C5, C3
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_027'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Break-Even' sheet must exist
    if 'Break-Even' not in wb.sheetnames:
        print(f"FAIL: 'Break-Even' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Break-Even']

    # Read the relevant cell values
    c2_val = ws['C2'].value  # Unit Price (should be 50)
    c3_val = ws['C3'].value  # Units Sold (should be 1250 after Goal Seek)
    c4_val = ws['C4'].value  # Variable Cost per Unit (should be 30)
    c5_val = ws['C5'].value  # Fixed Costs (should be 25000)

    print(f"DEBUG: C2 (Unit Price)={c2_val}, C3 (Units Sold)={c3_val}, "
          f"C4 (Var Cost)={c4_val}, C5 (Fixed Costs)={c5_val}")

    # Component 1: C3 contains the break-even value 1250 (0.6 points)
    # The break-even volume is 25000 / (50 - 30) = 1250
    # Goal Seek should have changed C3 from 800 to 1250
    try:
        if c3_val is not None:
            c3_numeric = float(c3_val)
            # Allow small tolerance for floating-point from Goal Seek
            if abs(c3_numeric - 1250.0) < 0.5:
                print(f"PASS: Component 1 — C3 contains {c3_numeric}, expected ~1250 (0.6 pts)")
                total_score += 0.6
            else:
                print(f"FAIL: Component 1 — C3 contains {c3_numeric}, expected ~1250")
        else:
            print("FAIL: Component 1 — C3 is None/empty")
    except (ValueError, TypeError) as e:
        # C3 might be a formula string if Goal Seek wasn't applied
        print(f"FAIL: Component 1 — C3 value '{c3_val}' is not numeric: {e}")

    # Component 2: Net income computes to ~0 at the break-even point (0.4 points)
    # Net Income = (Unit Price - Variable Cost) * Units Sold - Fixed Costs
    # At break-even: (50 - 30) * 1250 - 25000 = 0
    # We compute this from the actual cell values to verify mathematical consistency
    try:
        if all(v is not None for v in [c2_val, c3_val, c4_val, c5_val]):
            unit_price = float(c2_val)
            units_sold = float(c3_val)
            var_cost = float(c4_val)
            fixed_costs = float(c5_val)
            net_income = (unit_price - var_cost) * units_sold - fixed_costs
            print(f"DEBUG: Computed net income = ({unit_price} - {var_cost}) * {units_sold} - {fixed_costs} = {net_income}")
            # Allow tolerance for floating-point
            if abs(net_income) < 1.0:
                print(f"PASS: Component 2 — Net income is {net_income}, approximately 0 (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — Net income is {net_income}, expected ~0")
        else:
            print("FAIL: Component 2 — One or more required cells (C2-C5) are empty")
    except (ValueError, TypeError) as e:
        print(f"ERROR: Component 2 — Could not compute net income: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point — test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
