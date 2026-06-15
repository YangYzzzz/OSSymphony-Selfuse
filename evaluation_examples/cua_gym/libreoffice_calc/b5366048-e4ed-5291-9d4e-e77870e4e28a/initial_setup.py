"""
Initial Setup: Break-Even Analysis with Goal Seek
Task ID: calc_gg3_027
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Break-Even Sheet ---
    ws = wb.active
    ws.title = 'Break-Even'

    # Title row
    ws['A1'] = 'Break-Even Analysis'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Labels in column B, values in column C
    labels = {
        2: 'Unit Price',
        3: 'Units Sold',
        4: 'Variable Cost per Unit',
        5: 'Fixed Costs',
        6: 'Total Revenue',
        7: 'Net Income',
    }
    for row, label in labels.items():
        ws.cell(row=row, column=2, value=label)
        ws.cell(row=row, column=2).font = Font(name='Arial', size=11)

    # Values in column C
    ws['C2'] = 50        # Unit Price ($50)
    ws['C3'] = 800       # Units Sold (currently 800)
    ws['C4'] = 30        # Variable Cost per Unit ($30)
    ws['C5'] = 25000     # Fixed Costs ($25,000)

    # Formulas
    ws['C6'] = '=C2*C3'           # Total Revenue
    ws['C7'] = '=(C2-C4)*C3-C5'  # Net Income

    # Format currency cells
    for row in [2, 4, 5, 6, 7]:
        ws.cell(row=row, column=3).number_format = '$#,##0'
    # Units Sold is plain integer
    ws['C3'].number_format = '#,##0'

    # Style the labels
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
    for row in range(2, 8):
        cell = ws.cell(row=row, column=2)
        cell.font = Font(name='Arial', size=11, bold=True)

    # Add a separator line before calculated fields
    thin = Side(style='thin', color='000000')
    for col in range(2, 4):
        ws.cell(row=6, column=col).border = Border(top=thin)

    # Column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 5

    # Add a second sheet with supporting info for complexity
    ws2 = wb.create_sheet('Assumptions')
    ws2['A1'] = 'Key Assumptions'
    ws2['A1'].font = Font(name='Arial', size=12, bold=True)
    assumptions = [
        ['Market Segment', 'Mid-range Consumer Electronics'],
        ['Product Line', 'Portable Bluetooth Speakers'],
        ['Analysis Period', 'Q2 2025'],
        ['Distribution Channel', 'Online Direct-to-Consumer'],
        ['Target Market Size', '15,000 units/quarter'],
        ['Current Production Capacity', '2,000 units/month'],
        ['Warehouse Location', 'Portland, OR'],
        ['Lead Time', '14 business days'],
        ['Payment Terms', 'Net 30'],
        ['Warranty Period', '12 months'],
    ]
    for r, (key, val) in enumerate(assumptions, 3):
        ws2.cell(row=r, column=1, value=key)
        ws2.cell(row=r, column=1).font = Font(name='Arial', size=11, bold=True)
        ws2.cell(row=r, column=2, value=val)
        ws2.cell(row=r, column=2).font = Font(name='Arial', size=11)
    ws2.column_dimensions['A'].width = 28
    ws2.column_dimensions['B'].width = 35

    # Add a third sheet with cost breakdown for added realism
    ws3 = wb.create_sheet('Cost Breakdown')
    ws3['A1'] = 'Variable Cost Components'
    ws3['A1'].font = Font(name='Arial', size=12, bold=True)
    cost_headers = ['Component', 'Cost per Unit']
    for c, h in enumerate(cost_headers, 1):
        ws3.cell(row=2, column=c, value=h)
        ws3.cell(row=2, column=c).font = Font(name='Arial', size=11, bold=True)
    cost_data = [
        ['Raw Materials', 12.50],
        ['Assembly Labor', 8.00],
        ['Packaging', 3.50],
        ['Shipping per Unit', 4.00],
        ['Quality Control', 2.00],
    ]
    for r, (comp, cost) in enumerate(cost_data, 3):
        ws3.cell(row=r, column=1, value=comp)
        ws3.cell(row=r, column=2, value=cost)
        ws3.cell(row=r, column=2).number_format = '$#,##0.00'
    total_row = 3 + len(cost_data)
    ws3.cell(row=total_row, column=1, value='Total Variable Cost')
    ws3.cell(row=total_row, column=1).font = Font(name='Arial', size=11, bold=True)
    ws3.cell(row=total_row, column=2, value=f'=SUM(B3:B{total_row-1})')
    ws3.cell(row=total_row, column=2).number_format = '$#,##0.00'
    ws3.cell(row=total_row, column=2).font = Font(name='Arial', size=11, bold=True)
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
