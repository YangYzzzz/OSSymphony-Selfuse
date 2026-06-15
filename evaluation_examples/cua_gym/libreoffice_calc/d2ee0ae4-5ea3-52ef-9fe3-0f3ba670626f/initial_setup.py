"""
Initial Setup: Shipping Cost Calculator - Rate tables and empty calculator
Task ID: calc_wf_059
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Style definitions
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    money_fmt = '$#,##0.00'

    # Weight brackets used across all rate sheets
    weight_brackets = ["0-1 lb", "1-5 lb", "5-10 lb", "10-20 lb", "20-50 lb", "50+ lb"]
    zone_headers = ["Zone 1", "Zone 2", "Zone 3", "Zone 4", "Zone 5"]

    # --- UPS Rates ---
    ups_rates = [
        [4.50, 5.25, 6.80, 8.15, 10.50],
        [7.20, 8.90, 11.40, 14.25, 17.80],
        [10.85, 13.50, 17.20, 21.60, 26.40],
        [15.40, 19.75, 24.80, 30.50, 37.20],
        [22.60, 28.90, 36.50, 44.80, 54.30],
        [35.80, 44.50, 56.20, 68.90, 83.60],
    ]

    # --- FedEx Rates ---
    fedex_rates = [
        [4.25, 5.00, 6.50, 7.90, 9.80],
        [6.80, 8.50, 10.90, 13.60, 16.90],
        [10.20, 12.80, 16.50, 20.40, 25.10],
        [14.80, 18.90, 23.60, 29.20, 35.80],
        [21.50, 27.40, 34.80, 42.60, 52.00],
        [34.20, 42.80, 54.00, 66.50, 80.90],
    ]

    # --- USPS Rates ---
    usps_rates = [
        [3.80, 4.60, 5.90, 7.30, 9.20],
        [6.50, 8.10, 10.20, 12.80, 15.90],
        [9.80, 12.20, 15.60, 19.40, 23.80],
        [14.20, 18.10, 22.80, 28.00, 34.50],
        [20.80, 26.50, 33.60, 41.20, 50.40],
        [33.50, 41.80, 52.60, 64.80, 78.90],
    ]

    all_rates = {
        "UPS Rates": ups_rates,
        "FedEx Rates": fedex_rates,
        "USPS Rates": usps_rates,
    }

    # Create rate sheets
    first = True
    for sheet_name, rates in all_rates.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        # Title row
        ws.merge_cells("A1:F1")
        ws["A1"] = f"{sheet_name.replace(' Rates', '')} Shipping Rates"
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F5496")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Headers row 3
        ws.cell(row=3, column=1, value="Weight Bracket")
        ws.cell(row=3, column=1).font = header_font
        ws.cell(row=3, column=1).fill = header_fill
        ws.cell(row=3, column=1).alignment = header_align
        ws.cell(row=3, column=1).border = thin_border

        for j, zh in enumerate(zone_headers, 2):
            c = ws.cell(row=3, column=j, value=zh)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

        # Data rows 4-9
        for i, bracket in enumerate(weight_brackets):
            r = 4 + i
            ws.cell(row=r, column=1, value=bracket)
            ws.cell(row=r, column=1).font = Font(bold=True)
            ws.cell(row=r, column=1).border = thin_border
            for j, rate_val in enumerate(rates[i], 2):
                c = ws.cell(row=r, column=j, value=rate_val)
                c.number_format = money_fmt
                c.border = thin_border
                c.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 18
        for col_letter in ["B", "C", "D", "E", "F"]:
            ws.column_dimensions[col_letter].width = 12

    # --- Delivery Days Sheet ---
    ws_dd = wb.create_sheet("Delivery Days")
    ws_dd.merge_cells("A1:F1")
    ws_dd["A1"] = "Estimated Delivery Days by Carrier and Zone"
    ws_dd["A1"].font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    ws_dd["A1"].alignment = Alignment(horizontal="center")

    dd_headers = ["Carrier", "Zone 1", "Zone 2", "Zone 3", "Zone 4", "Zone 5"]
    for j, h in enumerate(dd_headers, 1):
        c = ws_dd.cell(row=3, column=j, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    delivery_data = [
        ["UPS", 2, 3, 4, 5, 7],
        ["FedEx", 2, 3, 3, 5, 6],
        ["USPS", 3, 4, 5, 6, 8],
    ]
    for i, row_data in enumerate(delivery_data):
        r = 4 + i
        for j, val in enumerate(row_data, 1):
            c = ws_dd.cell(row=r, column=j, value=val)
            c.border = thin_border
            c.alignment = Alignment(horizontal="center")
            if j == 1:
                c.font = Font(bold=True)

    ws_dd.column_dimensions["A"].width = 14
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws_dd.column_dimensions[col_letter].width = 12

    # --- Calculator Sheet ---
    ws_calc = wb.create_sheet("Calculator")

    # Title
    ws_calc.merge_cells("A1:F1")
    ws_calc["A1"] = "Shipping Cost Calculator"
    ws_calc["A1"].font = Font(name="Calibri", size=16, bold=True, color="2F5496")
    ws_calc["A1"].alignment = Alignment(horizontal="center")

    # Input section
    input_fill = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    ws_calc["A3"] = "Package Weight (lbs):"
    ws_calc["A3"].font = Font(bold=True, size=11)
    ws_calc["B3"] = 7.5
    ws_calc["B3"].font = Font(size=12)
    ws_calc["B3"].fill = input_fill
    ws_calc["B3"].border = thin_border
    ws_calc["B3"].number_format = '0.0'

    ws_calc["A4"] = "Destination Zone:"
    ws_calc["A4"].font = Font(bold=True, size=11)
    ws_calc["B4"] = 3
    ws_calc["B4"].font = Font(size=12)
    ws_calc["B4"].fill = input_fill
    ws_calc["B4"].border = thin_border

    # Results section header
    ws_calc.merge_cells("A6:E6")
    ws_calc["A6"] = "Rate Comparison Results"
    ws_calc["A6"].font = Font(name="Calibri", size=13, bold=True, color="FFFFFF")
    ws_calc["A6"].fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    ws_calc["A6"].alignment = Alignment(horizontal="center")

    result_headers = ["Carrier", "Shipping Rate", "Delivery Days", "Cost per Day"]
    for j, h in enumerate(result_headers, 1):
        c = ws_calc.cell(row=7, column=j, value=h)
        c.font = Font(bold=True, size=11)
        c.fill = PatternFill(start_color="FFB4C6E7", end_color="FFB4C6E7", fill_type="solid")
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    # Carrier labels in result rows (rows 8-10)
    carriers = ["UPS", "FedEx", "USPS"]
    for i, carrier in enumerate(carriers):
        r = 8 + i
        ws_calc.cell(row=r, column=1, value=carrier)
        ws_calc.cell(row=r, column=1).font = Font(bold=True, size=11)
        ws_calc.cell(row=r, column=1).border = thin_border
        # Leave columns B, C, D empty -- agent must fill with formulas
        for col in range(2, 5):
            ws_calc.cell(row=r, column=col).border = thin_border
            if col == 2:
                ws_calc.cell(row=r, column=col).number_format = money_fmt

    # Cheapest option label
    ws_calc["A12"] = "Cheapest Option:"
    ws_calc["A12"].font = Font(bold=True, size=12, color="006100")
    # B12 left empty -- agent fills with MIN formula

    ws_calc["A13"] = "Best Rate:"
    ws_calc["A13"].font = Font(bold=True, size=11)
    # B13 left empty -- agent fills with MIN formula

    # Column widths
    ws_calc.column_dimensions["A"].width = 24
    ws_calc.column_dimensions["B"].width = 16
    ws_calc.column_dimensions["C"].width = 16
    ws_calc.column_dimensions["D"].width = 16
    ws_calc.column_dimensions["E"].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
