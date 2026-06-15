"""
Reward Script: Shipping Cost Calculator
Task ID: calc_wf_059
Domain: libreoffice_calc
Scoring:
  Component 1: Rate lookup formulas (INDEX/MATCH) in B8:B10 — 0.30 pts
  Component 2: Delivery days lookup formulas in C8:C10 — 0.15 pts
  Component 3: Cost per day formulas in D8:D10 — 0.10 pts
  Component 4: Cheapest option formulas (MIN/INDEX) in B12:B13 — 0.15 pts
  Component 5: Conditional formatting on B8:B10 (green for cheapest) — 0.15 pts
  Component 6: Bar chart present on Calculator sheet — 0.15 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_059'


def persist_app_state(domain):
    """Send Ctrl+S to save any unsaved changes in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Calculator sheet must exist
    if 'Calculator' not in wb.sheetnames:
        print("CRITICAL: 'Calculator' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Calculator']

    # Component 1: Rate lookup formulas in B8:B10 using INDEX/MATCH (0.30 points)
    # These cells should contain INDEX(...MATCH...) formulas referencing the rate tables
    try:
        rate_formula_count = 0
        carrier_sheets = ["UPS Rates", "FedEx Rates", "USPS Rates"]
        for row_idx, expected_sheet in zip([8, 9, 10], carrier_sheets):
            cell_val = ws.cell(row=row_idx, column=2).value  # B8, B9, B10
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                formula_upper = cell_val.upper()
                # Must reference the carrier rate sheet and use INDEX or VLOOKUP
                if ('INDEX' in formula_upper or 'VLOOKUP' in formula_upper) and \
                   ('MATCH' in formula_upper or 'VLOOKUP' in formula_upper):
                    # Check that it references the correct carrier rate sheet
                    if expected_sheet.upper().replace(' ', '') in cell_val.upper().replace(' ', '') or \
                       expected_sheet.upper() in cell_val.upper():
                        rate_formula_count += 1
                        print(f"PASS: B{row_idx} has rate lookup formula referencing '{expected_sheet}'")
                    else:
                        print(f"PARTIAL: B{row_idx} has lookup formula but may not reference '{expected_sheet}': {cell_val}")
                        rate_formula_count += 0.5
                else:
                    print(f"FAIL: B{row_idx} has formula but not INDEX/MATCH or VLOOKUP: {cell_val}")
            else:
                print(f"FAIL: B{row_idx} has no formula, found: {cell_val}")

        if rate_formula_count >= 3:
            print(f"PASS: Component 1 - All 3 rate lookup formulas present (0.30 pts)")
            total_score += 0.30
        elif rate_formula_count >= 2:
            print(f"PARTIAL: Component 1 - {rate_formula_count}/3 rate lookup formulas (0.20 pts)")
            total_score += 0.20
        elif rate_formula_count >= 1:
            print(f"PARTIAL: Component 1 - {rate_formula_count}/3 rate lookup formulas (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 - No rate lookup formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Delivery days lookup formulas in C8:C10 (0.15 points)
    # These should reference the 'Delivery Days' sheet
    try:
        days_formula_count = 0
        for row_idx in [8, 9, 10]:
            cell_val = ws.cell(row=row_idx, column=3).value  # C8, C9, C10
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                formula_upper = cell_val.upper()
                if ('INDEX' in formula_upper or 'VLOOKUP' in formula_upper) and \
                   'DELIVERY' in formula_upper:
                    days_formula_count += 1
                    print(f"PASS: C{row_idx} has delivery days lookup formula")
                else:
                    print(f"FAIL: C{row_idx} has formula but doesn't reference Delivery Days: {cell_val}")
            else:
                print(f"FAIL: C{row_idx} has no formula, found: {cell_val}")

        if days_formula_count >= 3:
            print(f"PASS: Component 2 - All 3 delivery days formulas present (0.15 pts)")
            total_score += 0.15
        elif days_formula_count >= 2:
            print(f"PARTIAL: Component 2 - {days_formula_count}/3 delivery days formulas (0.10 pts)")
            total_score += 0.10
        elif days_formula_count >= 1:
            print(f"PARTIAL: Component 2 - {days_formula_count}/3 delivery days formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 - No delivery days formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Cost per day formulas in D8:D10 (0.10 points)
    # These should divide rate by days (B/C)
    try:
        cpd_formula_count = 0
        for row_idx in [8, 9, 10]:
            cell_val = ws.cell(row=row_idx, column=4).value  # D8, D9, D10
            if cell_val and isinstance(cell_val, str) and cell_val.startswith('='):
                formula_upper = cell_val.upper()
                # Should reference B and C columns in same row (division)
                if ('/' in cell_val or 'IF' in formula_upper):
                    cpd_formula_count += 1
                    print(f"PASS: D{row_idx} has cost per day formula: {cell_val}")
                else:
                    print(f"FAIL: D{row_idx} formula doesn't appear to compute cost per day: {cell_val}")
            else:
                print(f"FAIL: D{row_idx} has no formula, found: {cell_val}")

        if cpd_formula_count >= 3:
            print(f"PASS: Component 3 - All 3 cost per day formulas present (0.10 pts)")
            total_score += 0.10
        elif cpd_formula_count >= 2:
            print(f"PARTIAL: Component 3 - {cpd_formula_count}/3 cost per day formulas (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 - Insufficient cost per day formulas")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Cheapest option formulas in B12:B13 (0.15 points)
    # B12 should identify the cheapest carrier name, B13 should show the MIN rate
    try:
        cheapest_score = 0.0

        # B13: MIN formula for best rate
        b13_val = ws.cell(row=13, column=2).value
        if b13_val and isinstance(b13_val, str) and b13_val.startswith('='):
            if 'MIN' in b13_val.upper():
                cheapest_score += 0.075
                print(f"PASS: B13 has MIN formula: {b13_val}")
            else:
                print(f"FAIL: B13 formula doesn't use MIN: {b13_val}")
        else:
            print(f"FAIL: B13 has no formula, found: {b13_val}")

        # B12: INDEX/MATCH to identify cheapest carrier name
        b12_val = ws.cell(row=12, column=2).value
        if b12_val and isinstance(b12_val, str) and b12_val.startswith('='):
            formula_upper = b12_val.upper()
            if ('INDEX' in formula_upper or 'VLOOKUP' in formula_upper) and \
               ('MATCH' in formula_upper or 'MIN' in formula_upper):
                cheapest_score += 0.075
                print(f"PASS: B12 has cheapest carrier lookup formula: {b12_val}")
            else:
                print(f"FAIL: B12 formula doesn't use INDEX/MATCH with MIN: {b12_val}")
        else:
            print(f"FAIL: B12 has no formula, found: {b12_val}")

        if cheapest_score > 0:
            print(f"PASS: Component 4 - Cheapest option formulas ({cheapest_score} pts)")
            total_score += cheapest_score
        else:
            print(f"FAIL: Component 4 - No cheapest option formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting on B8:B10 highlighting cheapest (0.15 points)
    # Should have green conditional formatting applied to the rate cells
    try:
        cf_rules = list(ws.conditional_formatting)
        # Look for conditional formatting rules covering B8, B9, or B10
        cf_on_rates = 0
        for cf in cf_rules:
            cf_range_str = str(cf)
            for target in ['B8', 'B9', 'B10']:
                if target.lower() in cf_range_str.lower() or target in cf_range_str:
                    # Check if it's related to MIN comparison
                    for rule in cf.rules:
                        formula_str = ' '.join(str(f) for f in rule.formula) if rule.formula else ''
                        if 'MIN' in formula_str.upper() or rule.type == 'expression':
                            cf_on_rates += 1
                            break

        if cf_on_rates >= 3:
            print(f"PASS: Component 5 - Conditional formatting on all 3 rate cells (0.15 pts)")
            total_score += 0.15
        elif cf_on_rates >= 2:
            print(f"PARTIAL: Component 5 - Conditional formatting on {cf_on_rates}/3 rate cells (0.10 pts)")
            total_score += 0.10
        elif cf_on_rates >= 1:
            print(f"PARTIAL: Component 5 - Conditional formatting on {cf_on_rates}/3 rate cells (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 - No conditional formatting found on rate cells")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Bar chart present on Calculator sheet (0.15 points)
    # Should be a bar/column chart comparing carrier rates
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            chart_type = chart.__class__.__name__
            has_series = len(chart.series) >= 1

            if ('Bar' in chart_type or 'bar' in chart_type) and has_series:
                print(f"PASS: Component 6 - Bar chart found with {len(chart.series)} series (0.15 pts)")
                total_score += 0.15
            elif has_series:
                # Other chart type but still a comparison chart
                print(f"PARTIAL: Component 6 - Chart found ({chart_type}) with {len(chart.series)} series (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 6 - Chart found but no data series")
        else:
            print(f"FAIL: Component 6 - No charts found on Calculator sheet")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
