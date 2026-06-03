"""
Initial Setup: Create a production schedule table with daily output targets for 5 products over a work week.
Task ID: calc_ops_025
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'ProdSchedule'

    # Headers in row 1
    headers = ['Product', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Product Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Product data rows 2-6
    data = [
        ['Prod-A', 100, 120, 110, 100, 90],
        ['Prod-B', 50, 60, 55, 65, 50],
        ['Prod-C', 200, 180, 210, 190, 200],
        ['Prod-D', 75, 80, 70, 85, 75],
        ['Prod-E', 150, 140, 160, 155, 145],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Row 7 label
    ws.cell(row=7, column=1, value='Daily Total')

    # G2:G6 and B7:G7 are intentionally left EMPTY - the task is to add SUM formulas there

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
