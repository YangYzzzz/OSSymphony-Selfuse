"""
Reward Script: Production schedule table with SUM formulas for product totals and daily totals
Task ID: calc_ops_025
Domain: libreoffice_calc
Scoring:
  Component 1 (0.50): Product Total SUM formulas in G2:G6
  Component 2 (0.35): Daily Total SUM formulas in B7:F7
  Component 3 (0.15): Grand Total SUM formula in G7
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_025'


def check_formula(ws, coord, expected_formula):
    """Check if cell contains expected formula (case-insensitive, whitespace-stripped)."""
    val = ws[coord].value
    if not isinstance(val, str):
        return False, val
    return val.upper().replace(" ", "") == expected_formula.upper().replace(" ", ""), val


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'ProdSchedule' not in wb.sheetnames:
        print("FAIL: Sheet 'ProdSchedule' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ProdSchedule']

    # Component 1: Product Total SUM formulas in G2:G6 (0.50 points, 0.10 each)
    # These cells should contain =SUM(B<row>:F<row>) for each product row
    print("--- Component 1: Product Total formulas (G2:G6) ---")
    comp1_score = 0.0
    product_formulas = {
        'G2': '=SUM(B2:F2)',
        'G3': '=SUM(B3:F3)',
        'G4': '=SUM(B4:F4)',
        'G5': '=SUM(B5:F5)',
        'G6': '=SUM(B6:F6)',
    }
    for coord, expected in product_formulas.items():
        try:
            match, actual = check_formula(ws, coord, expected)
            if match:
                print(f"  PASS: {coord} contains {expected} (0.10 pts)")
                comp1_score += 0.10
            else:
                print(f"  FAIL: {coord} expected {expected}, found: {actual}")
        except Exception as e:
            print(f"  ERROR: {coord} — {e}")

    total_score += comp1_score
    print(f"  Component 1 subtotal: {comp1_score:.2f}/0.50")

    # Component 2: Daily Total SUM formulas in B7:F7 (0.35 points, 0.07 each)
    # These cells should contain =SUM(<col>2:<col>6) for each weekday column
    print("--- Component 2: Daily Total formulas (B7:F7) ---")
    comp2_score = 0.0
    daily_formulas = {
        'B7': '=SUM(B2:B6)',
        'C7': '=SUM(C2:C6)',
        'D7': '=SUM(D2:D6)',
        'E7': '=SUM(E2:E6)',
        'F7': '=SUM(F2:F6)',
    }
    for coord, expected in daily_formulas.items():
        try:
            match, actual = check_formula(ws, coord, expected)
            if match:
                print(f"  PASS: {coord} contains {expected} (0.07 pts)")
                comp2_score += 0.07
            else:
                print(f"  FAIL: {coord} expected {expected}, found: {actual}")
        except Exception as e:
            print(f"  ERROR: {coord} — {e}")

    total_score += comp2_score
    print(f"  Component 2 subtotal: {comp2_score:.2f}/0.35")

    # Component 3: Grand Total formula in G7 (0.15 points)
    # G7 should contain =SUM(G2:G6) (or equivalently =SUM(B7:F7))
    print("--- Component 3: Grand Total formula (G7) ---")
    try:
        val = ws['G7'].value
        if isinstance(val, str):
            norm = val.upper().replace(" ", "")
            # Accept either =SUM(G2:G6) or =SUM(B7:F7) — both produce 2915
            if norm == '=SUM(G2:G6)' or norm == '=SUM(B7:F7)':
                print(f"  PASS: G7 contains {val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"  FAIL: G7 expected =SUM(G2:G6) or =SUM(B7:F7), found: {val}")
        else:
            print(f"  FAIL: G7 expected a SUM formula, found: {val}")
    except Exception as e:
        print(f"  ERROR: G7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
