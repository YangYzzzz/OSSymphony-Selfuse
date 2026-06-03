"""
Initial Setup: Bio papers reference spreadsheet
Task ID: osworld_multi_apps_pdf_download_cite_008
Domain: libreoffice_calc (multi-app: also Chrome)

Creates:
  - /home/user/Desktop/bio_papers.xlsx  — spreadsheet with biology paper titles and PubMed URLs
  - Chrome is opened (already present in VM, just ensure it's running)
  - LibreOffice Calc opens bio_papers.xlsx

NOTE: No bio_paper01.pdf and no bio_citation_answer.docx in initial state
      (agent must create them as the task outcome)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_pdf_download_cite_008'
OUTPUT = f'{DESKTOP}/bio_papers.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(DESKTOP, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Biology Papers"

    # --- Column widths ---
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 45

    # --- Header row ---
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF2E75B6', end_color='FF2E75B6', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 24

    headers = ['Title', 'PubMed URL']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Paper data ---
    # Row 2: First paper — Robbins & Monro 1951 (foundational stochastic approximation)
    # Row 3: Kiefer & Wolfowitz 1952 — direct follow-up that cites Robbins & Monro 1951
    # Row 4: Dempster, Laird & Rubin 1977 — EM algorithm (classical stats)
    # Row 5: Cox 1972 — Proportional hazards (classical biostatistics)
    papers = [
        (
            'A Stochastic Approximation Method',
            'https://pubmed.ncbi.nlm.nih.gov/14783122/'
        ),
        (
            'Stochastic Estimation of the Maximum of a Regression Function',
            'https://pubmed.ncbi.nlm.nih.gov/12371600/'
        ),
        (
            'Maximum Likelihood from Incomplete Data via the EM Algorithm',
            'https://pubmed.ncbi.nlm.nih.gov/1244801/'
        ),
        (
            'Regression Models and Life-Tables',
            'https://pubmed.ncbi.nlm.nih.gov/17571013/'
        ),
    ]

    row_fill_light = PatternFill(start_color='FFDCE6F1', end_color='FFDCE6F1', fill_type='solid')
    row_fill_white = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
    data_font = Font(name='Calibri', size=11)
    url_font = Font(name='Calibri', size=11, color='FF0563C1', underline='single')
    thin = Side(style='thin', color='FFB8CCE4')

    for i, (title, url) in enumerate(papers, 2):
        fill = row_fill_light if i % 2 == 0 else row_fill_white

        title_cell = ws.cell(row=i, column=1, value=title)
        title_cell.font = data_font
        title_cell.fill = fill
        title_cell.alignment = Alignment(vertical='center', wrap_text=True)
        title_cell.border = Border(
            left=Side(style='thin', color='FFB8CCE4'),
            right=Side(style='thin', color='FFB8CCE4'),
            top=Side(style='thin', color='FFB8CCE4'),
            bottom=Side(style='thin', color='FFB8CCE4'),
        )

        url_cell = ws.cell(row=i, column=2, value=url)
        url_cell.font = url_font
        url_cell.fill = fill
        url_cell.alignment = Alignment(vertical='center', wrap_text=True)
        url_cell.border = Border(
            left=Side(style='thin', color='FFB8CCE4'),
            right=Side(style='thin', color='FFB8CCE4'),
            top=Side(style='thin', color='FFB8CCE4'),
            bottom=Side(style='thin', color='FFB8CCE4'),
        )
        ws.row_dimensions[i].height = 20

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Ensure bio_paper01.pdf and bio_citation_answer.docx do NOT exist
    for fname in ['bio_paper01.pdf', 'bio_citation_answer.docx']:
        fpath = os.path.join(WORKDIR, fname)
        if os.path.exists(fpath):
            os.remove(fpath)
            print(f'Removed pre-existing file: {fpath}')

    # GUI-ready startup: open LibreOffice Calc with the spreadsheet and Chrome
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui('google-chrome', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and Chrome with DISPLAY=:0')


create_initial()
