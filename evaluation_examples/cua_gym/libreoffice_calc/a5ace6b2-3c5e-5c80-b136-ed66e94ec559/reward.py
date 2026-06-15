"""
Reward Script: Apply custom number format '00000' to ZIP code column
Task ID: calc_fmt_numfmt_zip_code_082
Domain: libreoffice_calc
Scoring:
  Component 1: All cells D2:D60 have number format '00000' (0.7 pts)
  Component 2: Numeric values in D2:D60 are preserved unchanged,
               verified as compound check (format applied + values numeric) (0.3 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_zip_code_082'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Apply custom number format '00000' to cells D2:D60
    in the 'Address Book' sheet, so ZIP codes display with
    leading zeros without changing the underlying numeric values.
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: required sheet must exist
    if 'Address Book' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Address Book' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Address Book']

    # Component 1: All cells D2:D60 must have number format '00000' (0.7 points)
    # FAILS on initial (all 'General'), PASSES on golden (all '00000')
    try:
        cells_with_correct_format = sum(
            1 for row in range(2, 61)
            if ws.cell(row=row, column=4).number_format == '00000'
        )
        total_zip_cells = 59  # rows 2 through 60

        if cells_with_correct_format == total_zip_cells:
            print(f"PASS: Component 1 — All {total_zip_cells} cells D2:D60 have '00000' format (0.7 pts)")
            total_score += 0.7
        elif cells_with_correct_format > 0:
            # Partial: some cells correctly formatted — task requires ALL
            wrong_cells = [
                f"D{row}: {repr(ws.cell(row=row, column=4).number_format)}"
                for row in range(2, 61)
                if ws.cell(row=row, column=4).number_format != '00000'
            ]
            fraction = cells_with_correct_format / total_zip_cells
            partial = round(0.7 * fraction, 2)
            if partial > 0:
                print(f"PARTIAL: Component 1 — {cells_with_correct_format}/{total_zip_cells} cells have '00000' format")
                print(f"  First wrong cells: {wrong_cells[:5]}")
                total_score += partial
        else:
            sample_fmt = ws.cell(row=2, column=4).number_format
            print(f"FAIL: Component 1 — No cells in D2:D60 have '00000' format (sample D2: {repr(sample_fmt)})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        cells_with_correct_format = 0

    # Component 2: ZIP values must remain numeric (not converted to text), compound check
    # with format applied as a gate. (0.3 points)
    # On initial file: numeric values exist but format='General' → compound FAILS
    # On golden file: numeric values exist AND format='00000' → compound PASSES
    try:
        numeric_count = sum(
            1 for row in range(2, 61)
            if isinstance(ws.cell(row=row, column=4).value, (int, float))
        )
        non_numeric = [
            f"D{row}: {repr(ws.cell(row=row, column=4).value)}"
            for row in range(2, 61)
            if not isinstance(ws.cell(row=row, column=4).value, (int, float))
        ]

        # Compound condition: ALL 59 values are numeric AND format '00000' was applied
        all_numeric = (numeric_count == 59)
        format_fully_applied = (cells_with_correct_format == 59)

        if all_numeric and format_fully_applied:
            print(f"PASS: Component 2 — All {numeric_count} ZIP values are numeric with '00000' format applied (0.3 pts)")
            total_score += 0.3
        elif all_numeric and not format_fully_applied:
            print(f"FAIL: Component 2 — Values are numeric but '00000' format not fully applied (initial state or partial)")
        elif not all_numeric:
            print(f"FAIL: Component 2 — Some ZIP values are not numeric: {non_numeric[:3]}")
        else:
            print(f"FAIL: Component 2 — {numeric_count}/59 numeric, {cells_with_correct_format}/59 correct format")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Info check: D1 header should remain 'ZIP' (sanity, no scoring)
    try:
        d1_val = ws['D1'].value
        if d1_val == 'ZIP':
            print(f"INFO: D1 header 'ZIP' intact (format: {repr(ws['D1'].number_format)})")
        else:
            print(f"WARN: D1 header unexpected value: {repr(d1_val)}")
    except Exception as e:
        print(f"INFO: Could not check D1: {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
