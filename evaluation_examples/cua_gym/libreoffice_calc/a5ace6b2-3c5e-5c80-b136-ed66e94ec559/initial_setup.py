"""
Initial Setup: Address Book with ZIP codes stored as numbers (leading zeros stripped)
Task ID: calc_fmt_numfmt_zip_code_082
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_zip_code_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Address Book ---
    ws = wb.active
    ws.title = 'Address Book'

    # Headers
    headers = ['Name', 'Address', 'City', 'ZIP', 'State']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic address data (59 rows, rows 2-60)
    # ZIP codes include some with leading zeros (stored as int, so leading zero lost)
    data = [
        ('Sarah Chen',        '142 Maple Street',      'Boston',        2134,  'MA'),
        ('Marcus Johnson',    '789 Oak Avenue',         'Newark',        7301,  'NJ'),
        ('Emily Rodriguez',   '4501 Sunset Blvd',       'Los Angeles',   90210, 'CA'),
        ('James Williams',    '33 Park Place',          'New York',      10001, 'NY'),
        ('Linda Nakamura',    '2200 Harbor View Rd',    'Seattle',       98101, 'WA'),
        ('David Kim',         '88 Peachtree Street',    'Atlanta',       30301, 'GA'),
        ('Priya Patel',       '6700 W Lake Shore Dr',   'Chicago',       60601, 'IL'),
        ('Thomas Anderson',   '5 Freedom Trail',        'Providence',    2901,  'RI'),
        ('Angela Torres',     '310 Mission Street',     'San Francisco', 94105, 'CA'),
        ('Robert Singh',      '1001 Congress Ave',      'Austin',        73301, 'TX'),
        ('Karen White',       '450 N Michigan Ave',     'Chicago',       60611, 'IL'),
        ('Michael Brown',     '2801 University Blvd',   'Albuquerque',   87106, 'NM'),
        ('Jessica Davis',     '17 Beacon Street',       'Boston',        2108,  'MA'),
        ('William Martinez',  '900 King Street',        'Charleston',    29403, 'SC'),
        ('Amanda Taylor',     '3300 Wilshire Blvd',     'Los Angeles',   90010, 'CA'),
        ('Christopher Lee',   '1540 Broadway',          'New York',      10036, 'NY'),
        ('Nicole Clark',      '250 Post Road',          'Westport',      6880,  'CT'),
        ('Daniel Lewis',      '1600 Pennsylvania Ave',  'Washington',    20500, 'DC'),
        ('Stephanie Hall',    '420 Central Park West',  'New York',      10025, 'NY'),
        ('Joshua Walker',     '700 Boylston Street',    'Boston',        2116,  'MA'),
        ('Rachel Allen',      '220 Canal Street',       'New Orleans',   70130, 'LA'),
        ('Kevin Young',       '38 Newbury Street',      'Boston',        2116,  'MA'),
        ('Megan Hernandez',   '1900 Lakeside Drive',    'Cleveland',     44114, 'OH'),
        ('Brian Scott',       '500 Market Street',      'Philadelphia',  19106, 'PA'),
        ('Heather Green',     '2100 Fairview Ave N',    'Seattle',       98109, 'WA'),
        ('Anthony Nelson',    '600 Travis Street',      'Houston',       77002, 'TX'),
        ('Michelle Carter',   '808 Nuuanu Avenue',      'Honolulu',      96817, 'HI'),
        ('Jason Mitchell',    '3600 Wilshire Blvd',     'Los Angeles',   90010, 'CA'),
        ('Samantha Perez',    '45 West 57th Street',    'New York',      10019, 'NY'),
        ('Ryan Roberts',      '15 Stearns Wharf',       'Santa Barbara', 93101, 'CA'),
        ('Brittany Turner',   '100 Grand Avenue',       'Oakland',       94612, 'CA'),
        ('Andrew Phillips',   '700 Arch Street',        'Philadelphia',  19106, 'PA'),
        ('Lauren Campbell',   '1 Maritime Plaza',       'San Francisco', 94111, 'CA'),
        ('Justin Parker',     '333 Wacker Drive',       'Chicago',       60606, 'IL'),
        ('Vanessa Evans',     '200 S Biscayne Blvd',    'Miami',         33131, 'FL'),
        ('Aaron Edwards',     '80 Pine Street',         'New York',      10005, 'NY'),
        ('Tiffany Collins',   '600 Grant Street',       'Pittsburgh',    15219, 'PA'),
        ('Brandon Stewart',   '1500 Market Street',     'Philadelphia',  19102, 'PA'),
        ('Amber Sanchez',     '7 Willow Lane',          'Greenwich',     6830,  'CT'),
        ('Patrick Morris',    '330 Front Street',       'Columbus',      43215, 'OH'),
        ('Danielle Rogers',   '1200 Elm Street',        'Manchester',    3101,  'NH'),
        ('Timothy Reed',      '25 Boylston Street',     'Boston',        2116,  'MA'),
        ('Crystal Cook',      '900 Wabasha Street',     'Saint Paul',    55102, 'MN'),
        ('Steven Morgan',     '1400 Canal Street',      'New Orleans',   70112, 'LA'),
        ('Elizabeth Bell',    '400 Market Street',      'San Francisco', 94111, 'CA'),
        ('Gregory Murphy',    '600 Burrard Street',     'Vancouver',     2172,  'WA'),
        ('Natasha Bailey',    '2500 Lincoln Park West', 'Chicago',       60614, 'IL'),
        ('Matthew Rivera',    '1700 Broadway',          'New York',      10019, 'NY'),
        ('Kimberly Cooper',   '4400 Vail Road',         'Denver',        80205, 'CO'),
        ('Tyler Richardson',  '650 Bush Street',        'San Francisco', 94108, 'CA'),
        ('Alexis Cox',        '1 Faneuil Hall Square',  'Boston',        2109,  'MA'),
        ('Jordan Howard',     '2200 M Street NW',       'Washington',    20037, 'DC'),
        ('Cassandra Ward',    '480 Lexington Ave',      'New York',      10017, 'NY'),
        ('Nathan Torres',     '330 S Hope Street',      'Los Angeles',   90071, 'CA'),
        ('Jasmine Peterson',  '111 S Michigan Ave',     'Chicago',       60603, 'IL'),
        ('Trevor Gray',       '34 Federal Street',      'Boston',        2110,  'MA'),
        ('Shannon Ramirez',   '6400 Sunset Blvd',       'Los Angeles',   90028, 'CA'),
        ('Zachary James',     '210 Post Street',        'San Francisco', 94108, 'CA'),
        ('Monica Watson',     '90 Church Street',       'Burlington',    5401,  'VT'),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column D (ZIP) is stored as numeric with General format (default)
    # Do NOT apply '00000' format — that is what the task asks the agent to do

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Rows written: {len(data)} data rows (rows 2-{len(data)+1})')
    print('Column D ZIP codes stored as plain integers with General format')


create_initial()
