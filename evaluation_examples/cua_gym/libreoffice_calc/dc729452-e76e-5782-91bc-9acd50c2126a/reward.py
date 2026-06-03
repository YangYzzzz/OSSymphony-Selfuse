"""
Reward Script: Data audit on product database - summarize issues on AuditResults sheet
Task ID: calc_gen_data_cleanup_060
Domain: libreoffice_calc

Scoring rubric (total = 1.0):
  Component 1: Products sheet has column G 'Issues' header + flags for issue rows (0.25 pts)
  Component 2: AuditResults Section 1 has correct count (8) and >=7 sell<cost products listed (0.25 pts)
  Component 3: AuditResults Section 2 has correct count (12) and >=11 zero/neg stock products listed (0.25 pts)
  Component 4: AuditResults Section 3 has correct count (15) and >=14 invalid-code products listed (0.25 pts)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_data_cleanup_060'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both sheets must exist
    if 'Products' not in wb.sheetnames:
        print("CRITICAL: 'Products' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'AuditResults' not in wb.sheetnames:
        print("CRITICAL: 'AuditResults' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_prod = wb['Products']
    ws_audit = wb['AuditResults']

    # -----------------------------------------------------------------------
    # Component 1: Column G 'Issues' flags in Products sheet (0.25 points)
    # The initial file has NO column G data (all None). The golden file has
    # "Issues" as the G1 header and non-empty flag strings for all issue rows.
    # We require:
    #   a) G1 header is non-empty (some label indicating issues column)
    #   b) At least 20 rows with non-empty G flags (golden has 31)
    #   c) At least one flag value contains a known issue keyword
    # -----------------------------------------------------------------------
    try:
        g1_header = ws_prod.cell(row=1, column=7).value
        # Count non-empty G column values in data rows
        g_non_empty = 0
        g_has_known_flags = False
        for row in range(2, 302):
            val = ws_prod.cell(row=row, column=7).value
            if val:
                g_non_empty += 1
                val_str = str(val).upper()
                if any(kw in val_str for kw in ['PRICE', 'STOCK', 'CODE', 'ERROR']):
                    g_has_known_flags = True

        if g1_header and g_non_empty >= 20 and g_has_known_flags:
            print(f"PASS: Component 1 — Column G has header '{g1_header}', {g_non_empty} flagged rows with known issue keywords (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — G1 header='{g1_header}', flagged rows={g_non_empty}, has_known_flags={g_has_known_flags}")
            print(f"  Expected: non-empty G1 header, >=20 flagged rows, with PRICE/STOCK/CODE keywords")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Collect audit rows for section analysis (used in Components 2, 3, 4)
    # -----------------------------------------------------------------------
    audit_rows = []
    for row in ws_audit.iter_rows(values_only=True):
        audit_rows.append(row)

    # Find section boundaries by looking for "SECTION N" prefix in first cell
    # This is more reliable than keyword matching that can false-positive on data
    sec1_start = None
    sec2_start = None
    sec3_start = None

    for i, row in enumerate(audit_rows):
        first_cell = str(row[0]).upper() if row[0] is not None else ''
        if sec1_start is None and 'SECTION 1' in first_cell:
            sec1_start = i
        if sec2_start is None and 'SECTION 2' in first_cell:
            sec2_start = i
        if sec3_start is None and 'SECTION 3' in first_cell:
            sec3_start = i

    # Fallback: if SECTION N headers aren't found, try detecting by sell/stock/invalid keywords
    # but ONLY in the first cell and only for specific phrases
    if sec1_start is None:
        for i, row in enumerate(audit_rows):
            first_cell = str(row[0]).upper() if row[0] is not None else ''
            if 'SELL' in first_cell and 'COST' in first_cell:
                sec1_start = i
                break
    if sec2_start is None:
        for i, row in enumerate(audit_rows):
            first_cell = str(row[0]).upper() if row[0] is not None else ''
            if ('ZERO' in first_cell or 'NEGATIVE' in first_cell) and 'STOCK' in first_cell:
                sec2_start = i
                break
    if sec3_start is None:
        for i, row in enumerate(audit_rows):
            first_cell = str(row[0]).upper() if row[0] is not None else ''
            if 'INVALID' in first_cell and ('CODE' in first_cell or 'FORMAT' in first_cell):
                sec3_start = i
                break

    print(f"DEBUG: sec1_start={sec1_start}, sec2_start={sec2_start}, sec3_start={sec3_start}")

    # -----------------------------------------------------------------------
    # Component 2: AuditResults Section 1 — products with sell < cost (0.25 points)
    # Golden has "Issue Count: 8" and 8 product rows.
    # -----------------------------------------------------------------------
    try:
        # Search for "Issue Count: 8" anywhere in the sheet
        sec1_count_found = False
        sec1_count_value = None
        for i, row in enumerate(audit_rows):
            for cell_val in row:
                if cell_val is not None:
                    val_str = str(cell_val)
                    if re.search(r'(?:issue\s*count|count)\s*[:\-]?\s*8\b', val_str, re.IGNORECASE):
                        sec1_count_found = True
                        sec1_count_value = val_str
                        break

        # Count product rows in section 1 (rows between sec1_start and sec2_start,
        # skipping section header, count row, and column header row)
        sec1_products = 0
        if sec1_start is not None and sec2_start is not None and sec2_start > sec1_start:
            for i in range(sec1_start + 3, sec2_start):
                row = audit_rows[i]
                first_val = row[0]
                if first_val is not None and str(first_val).strip():
                    first_upper = str(first_val).strip().upper()
                    # Skip column header rows
                    if first_upper not in ('PRODUCT CODE', 'CODE', 'NAME'):
                        sec1_products += 1

        if sec1_count_found and sec1_products >= 7:
            print(f"PASS: Component 2 — Section 1 'Issue Count: 8' found ('{sec1_count_value}'), {sec1_products} product rows (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — sec1_count_found={sec1_count_found} (value='{sec1_count_value}'), sec1_products={sec1_products}")
            print(f"  Expected: 'Issue Count: 8' in audit sheet, and >=7 product rows in Section 1")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: AuditResults Section 2 — zero/negative stock (0.25 points)
    # Golden has "Issue Count: 12" and 12 product rows.
    # -----------------------------------------------------------------------
    try:
        sec2_count_found = False
        sec2_count_value = None
        for i, row in enumerate(audit_rows):
            for cell_val in row:
                if cell_val is not None:
                    val_str = str(cell_val)
                    if re.search(r'(?:issue\s*count|count)\s*[:\-]?\s*12\b', val_str, re.IGNORECASE):
                        sec2_count_found = True
                        sec2_count_value = val_str
                        break

        sec2_products = 0
        if sec2_start is not None and sec3_start is not None and sec3_start > sec2_start:
            for i in range(sec2_start + 3, sec3_start):
                row = audit_rows[i]
                first_val = row[0]
                if first_val is not None and str(first_val).strip():
                    first_upper = str(first_val).strip().upper()
                    if first_upper not in ('PRODUCT CODE', 'CODE', 'NAME'):
                        sec2_products += 1

        if sec2_count_found and sec2_products >= 11:
            print(f"PASS: Component 3 — Section 2 'Issue Count: 12' found ('{sec2_count_value}'), {sec2_products} product rows (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — sec2_count_found={sec2_count_found} (value='{sec2_count_value}'), sec2_products={sec2_products}")
            print(f"  Expected: 'Issue Count: 12' in audit sheet, and >=11 product rows in Section 2")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: AuditResults Section 3 — invalid product code format (0.25 points)
    # Golden has "Issue Count: 15" and 15 product rows with invalid codes.
    # -----------------------------------------------------------------------
    try:
        sec3_count_found = False
        sec3_count_value = None
        for i, row in enumerate(audit_rows):
            for cell_val in row:
                if cell_val is not None:
                    val_str = str(cell_val)
                    if re.search(r'(?:issue\s*count|count)\s*[:\-]?\s*15\b', val_str, re.IGNORECASE):
                        sec3_count_found = True
                        sec3_count_value = val_str
                        break

        sec3_products = 0
        if sec3_start is not None:
            end_row = len(audit_rows)
            for i in range(sec3_start + 3, end_row):
                row = audit_rows[i]
                first_val = row[0]
                if first_val is not None and str(first_val).strip():
                    first_upper = str(first_val).strip().upper()
                    if first_upper not in ('PRODUCT CODE', 'CODE', 'NAME'):
                        sec3_products += 1

        if sec3_count_found and sec3_products >= 14:
            print(f"PASS: Component 4 — Section 3 'Issue Count: 15' found ('{sec3_count_value}'), {sec3_products} product rows (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 — sec3_count_found={sec3_count_found} (value='{sec3_count_value}'), sec3_products={sec3_products}")
            print(f"  Expected: 'Issue Count: 15' in audit sheet, and >=14 product rows in Section 3")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
