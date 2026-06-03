"""
Initial Setup: Product database audit task
Task ID: calc_gen_data_cleanup_060
Domain: libreoffice_calc

Creates a Products sheet with 300 rows containing ~8 sell<cost errors,
~12 zero/negative stock rows, ~15 invalid product code format rows,
and an empty AuditResults sheet.
"""

import openpyxl
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_data_cleanup_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

PRODUCT_CATEGORIES = [
    ('Electronics', ['Wireless Headphones', 'Bluetooth Speaker', 'USB-C Hub', 'Laptop Stand',
                      'Mechanical Keyboard', 'Ergonomic Mouse', 'Monitor Light Bar', 'Webcam HD',
                      'Noise Cancelling Earbuds', 'Smart Power Strip', 'Cable Management Kit',
                      'Portable SSD', 'Screen Cleaner Kit', 'Desk Lamp LED', 'Mini Projector']),
    ('Office', ['Stapler Premium', 'Paper Shredder', 'Label Printer', 'File Organizer',
                 'Whiteboard Markers', 'Desk Organizer', 'Letter Tray', 'Hanging Files',
                 'Binder Clips Pack', 'Index Tabs', 'Correction Tape', 'Post-It Notes Large',
                 'Rubber Bands Pack', 'Push Pins Box', 'Sticky Notes Pack']),
    ('Furniture', ['Adjustable Desk', 'Ergonomic Chair', 'Monitor Stand', 'Bookshelf Unit',
                    'Storage Cabinet', 'Drawer Organizer', 'Cable Raceway', 'Coat Hook Rail',
                    'Floating Shelf', 'Corner Desk', 'Mobile Pedestal', 'Filing Cabinet',
                    'Whiteboard Easel', 'Reception Desk', 'Conference Chair']),
    ('Peripherals', ['Trackball Mouse', 'Numeric Keypad', 'USB Hub 7-Port', 'KVM Switch',
                      'HDMI Splitter', 'Ethernet Switch', 'Docking Station', 'Graphics Tablet',
                      'Barcode Scanner', 'Card Reader USB', 'Fingerprint Reader', 'USB Microphone',
                      'Drawing Tablet', 'Conference Webcam', 'Smart Card Reader']),
    ('Supplies', ['Printer Paper A4', 'Toner Cartridge Black', 'Toner Cartridge Color',
                   'Laminator Pouches', 'Thermal Paper Roll', 'Name Badge Holders',
                   'Envelope Pack', 'Bubble Wrap Roll', 'Packing Tape', 'Box Cutters',
                   'Shipping Labels', 'Protective Sleeves', 'Stamp Pad', 'Rubber Stamp',
                   'Marking Tape']),
]

def generate_valid_code(n):
    """Generate valid XX-#### product code."""
    letters = 'ABCDEFGHJKLMNPQRSTUVWXYZ'
    prefix = letters[n // 1000 % len(letters)] + letters[n % len(letters)]
    num = 1000 + (n * 37 % 9000)
    return f'{prefix}-{num:04d}'

def generate_invalid_code(n):
    """Generate invalid product code (various bad formats)."""
    bad_formats = [
        f'X-{1000 + n % 9000}',           # only one letter
        f'XYZ-{1000 + n % 9000}',         # three letters
        f'AB-{10 + n % 900}',             # only 2-3 digits
        f'CD{1000 + n % 9000}',           # missing dash
        f'EF-{10000 + n % 90000}',        # 5 digits
        f'12-{1000 + n % 9000}',          # digits instead of letters
        f'GH-A{100 + n % 900}',           # letter in number part
        f'IJ {1000 + n % 9000}',          # space instead of dash
    ]
    return bad_formats[n % len(bad_formats)]

def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet 1: Products ----
    ws = wb.active
    ws.title = 'Products'

    headers = ['Product Code', 'Name', 'Cost Price', 'Sell Price', 'Stock Qty', 'Active']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Pre-decide which rows will have issues (deterministic with seed=42)
    total_rows = 300

    # ~8 rows with sell < cost (price inversion)
    sell_lt_cost_rows = sorted(random.sample(range(2, 301), 8))
    # ~12 rows with stock <= 0
    zero_stock_rows = sorted(random.sample(range(2, 301), 12))
    # ~15 rows with invalid product code (ensure some don't overlap with above for variety)
    invalid_code_rows = sorted(random.sample(range(2, 301), 15))

    sell_lt_cost_set = set(sell_lt_cost_rows)
    zero_stock_set = set(zero_stock_rows)
    invalid_code_set = set(invalid_code_rows)

    # Product name pool - cycle through categories
    all_names = []
    for category, names in PRODUCT_CATEGORIES:
        for name in names:
            all_names.append((category, name))

    invalid_code_counter = 0

    for row_idx in range(2, 302):
        i = row_idx - 2  # 0-indexed

        # Product code
        if row_idx in invalid_code_set:
            code = generate_invalid_code(invalid_code_counter)
            invalid_code_counter += 1
        else:
            code = generate_valid_code(i)

        # Product name (cycle through realistic names)
        cat_idx = i % len(all_names)
        _, name = all_names[cat_idx]
        # Add variation to avoid exact duplicates
        name_variants = ['', ' Pro', ' Plus', ' Lite', ' Mini', ' Max', ' Elite', ' Standard',
                         ' Premium', ' Basic', ' Advanced', ' Ultra', ' Compact', ' Deluxe', ' Series II']
        full_name = name + name_variants[i % len(name_variants)]

        # Cost price: $5.00 to $500.00
        cost = round(random.uniform(5, 500), 2)

        # Sell price: normally 10-40% above cost
        if row_idx in sell_lt_cost_set:
            # ERROR: sell < cost (multiply by 0.5 to 0.95)
            sell = round(cost * random.uniform(0.5, 0.95), 2)
        else:
            sell = round(cost * random.uniform(1.1, 1.4), 2)

        # Stock quantity
        if row_idx in zero_stock_set:
            # ERROR: zero or negative stock
            stock = random.choice([0, -1, -3, -5, -10, 0, 0, -2, -4, -8])
        else:
            stock = random.randint(5, 500)

        # Active flag
        active = random.choice(['Yes', 'Yes', 'Yes', 'No'])  # 75% active

        ws.cell(row=row_idx, column=1, value=code)
        ws.cell(row=row_idx, column=2, value=full_name)
        ws.cell(row=row_idx, column=3, value=cost)
        ws.cell(row=row_idx, column=4, value=sell)
        ws.cell(row=row_idx, column=5, value=stock)
        ws.cell(row=row_idx, column=6, value=active)

    # Column G must be empty (no flags yet - that's the task)

    # ---- Sheet 2: AuditResults (empty) ----
    ws2 = wb.create_sheet('AuditResults')
    # Leave completely empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Products sheet: 300 rows + 1 header')
    print(f'  AuditResults sheet: empty')
    print(f'  Sell<Cost rows: {sell_lt_cost_rows}')
    print(f'  Zero/Negative stock rows: {zero_stock_rows}')
    print(f'  Invalid code rows: {invalid_code_rows}')

create_initial()
