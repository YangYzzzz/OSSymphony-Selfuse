"""
Initial Setup: Apply custom number format to variance column in spreadsheet
Task ID: calc_fmt_numfmt_custom_027
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_custom_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Variance Report'

    # --- Headers ---
    headers = ['Account', 'Budget', 'Actual', 'Variance']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Data rows 2-25 (24 rows of realistic account data) ---
    # Mix of positive and negative integers in column D (Variance)
    data = [
        ('Revenue - Product Sales',   850000,  854500,  4500),
        ('Revenue - Service Fees',    125000,  122700, -2300),
        ('Revenue - Licensing',        48000,   49200,  1200),
        ('Revenue - Subscriptions',   320000,  311100, -8900),
        ('COGS - Materials',          210000,  207850,  2150),
        ('COGS - Labor',              180000,  183200, -3200),
        ('COGS - Overhead',            95000,   94100,   900),
        ('Operating - Salaries',      420000,  418600,  1400),
        ('Operating - Rent',           72000,   72000,     0),
        ('Operating - Utilities',      18000,   19350, -1350),
        ('Operating - Marketing',      65000,   60800,  4200),
        ('Operating - Travel',         24000,   27500, -3500),
        ('Operating - Software',       31000,   29650,  1350),
        ('Operating - Insurance',      15000,   14800,   200),
        ('Operating - Consulting',     42000,   46200, -4200),
        ('R&D - Personnel',           155000,  153400,  1600),
        ('R&D - Equipment',            28000,   30100, -2100),
        ('R&D - Supplies',             12000,   11500,   500),
        ('Admin - Legal',              19000,   21500, -2500),
        ('Admin - Accounting',         22000,   22000,     0),
        ('Admin - HR',                 16000,   14900,  1100),
        ('Depreciation - Equipment',   38000,   38000,     0),
        ('Interest Expense',           11000,   12400, -1400),
        ('Tax Provision',             142000,  139800,  2200),
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column D uses General format (no custom number format)
    # This is the default — do NOT apply any number format to D2:D25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
