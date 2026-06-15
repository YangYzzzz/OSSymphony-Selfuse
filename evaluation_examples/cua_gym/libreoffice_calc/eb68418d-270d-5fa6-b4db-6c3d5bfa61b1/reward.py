"""
Reward Script: Apply custom number format '#,##0;[RED]-#,##0' to cells D2:D25
Task ID: calc_fmt_numfmt_custom_027
Domain: libreoffice_calc
Scoring:
  Component 1 (0.7): All 24 cells D2:D25 have the correct custom number format
                     '#,##0;[RED]-#,##0' (FAILS on initial: all 'General')
  Component 2 (0.3): Correct format applied AND underlying values preserved AND
                     no other cells outside D2:D25 were reformatted
                     (compound check anchored to the format change)
"""

import os
import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_numfmt_custom_027'

# Expected custom format code (the only change from initial to golden)
EXPECTED_FORMAT = '#,##0;[RED]-#,##0'

# Expected values in D2:D25 (must remain unchanged after format application)
EXPECTED_D_VALUES = {
    2: 4500, 3: -2300, 4: 1200, 5: -8900, 6: 2150, 7: -3200,
    8: 900, 9: 1400, 10: 0, 11: -1350, 12: 4200, 13: -3500,
    14: 1350, 15: 200, 16: -4200, 17: 1600, 18: -2100, 19: 500,
    20: -2500, 21: 0, 22: 1100, 23: 0, 24: -1400, 25: 2200
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists (precondition gate — no points)
    if 'Variance Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Variance Report' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Variance Report']

    # -------------------------------------------------------------------------
    # Component 1: All 24 cells D2:D25 have format '#,##0;[RED]-#,##0' (0.7 pts)
    # Initial state: all D2:D25 use 'General' — FAILS on initial
    # Golden state: all D2:D25 use '#,##0;[RED]-#,##0' — PASSES on golden
    # -------------------------------------------------------------------------
    try:
        correctly_formatted = 0
        incorrectly_formatted = []

        for row in range(2, 26):  # rows 2 through 25 inclusive
            cell = ws.cell(row=row, column=4)  # column D
            actual_fmt = cell.number_format
            if actual_fmt == EXPECTED_FORMAT:
                correctly_formatted += 1
            else:
                incorrectly_formatted.append(f"D{row}: '{actual_fmt}'")

        if correctly_formatted == 24:
            print(f"PASS: Component 1 — All 24 cells D2:D25 have format '{EXPECTED_FORMAT}' (0.7 pts)")
            total_score += 0.7
        elif correctly_formatted > 0:
            # Partial credit proportional to number of correctly formatted cells
            partial = round((correctly_formatted / 24) * 0.7, 4)
            print(f"PARTIAL: Component 1 — {correctly_formatted}/24 cells have correct format ({partial} pts)")
            print(f"  Incorrectly formatted cells: {incorrectly_formatted[:5]}{'...' if len(incorrectly_formatted) > 5 else ''}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells have custom format. Sample: {incorrectly_formatted[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Custom format applied AND values unchanged AND no outside-scope
    #              cells were reformatted (0.3 pts)
    # This is a compound check: ONLY passes when at least some format has been applied
    # (correctly_formatted > 0 is required — anchored to the task change).
    # FAILS on initial because correctly_formatted == 0 there.
    # -------------------------------------------------------------------------
    try:
        # This component only applies when at least some of the target format was applied
        # (otherwise Component 1 already failed and we're looking at the wrong state)
        if correctly_formatted == 0:
            print("FAIL: Component 2 — Skipped: no format applied yet (requires Component 1 partial credit)")
        else:
            sub_issues = []

            # Sub-check A: Underlying values in D2:D25 are still correct
            for row, expected_val in EXPECTED_D_VALUES.items():
                cell = ws.cell(row=row, column=4)
                actual_val = cell.value
                if actual_val != expected_val:
                    sub_issues.append(f"D{row}: expected {expected_val}, found {repr(actual_val)}")

            # Sub-check B: D1 header not reformatted (out of scope)
            d1_fmt = ws.cell(row=1, column=4).number_format
            if d1_fmt not in ('General', '@', ''):
                sub_issues.append(f"D1 format incorrectly changed to '{d1_fmt}'")

            # Sub-check C: Columns A, B, C formats untouched (should remain General)
            for row in range(1, 26):
                for col in [1, 2, 3]:
                    fmt = ws.cell(row=row, column=col).number_format
                    coord = f'{get_column_letter(col)}{row}'
                    if fmt not in ('General', '@', ''):
                        sub_issues.append(f"{coord} format incorrectly changed to '{fmt}'")

            if len(sub_issues) == 0:
                print(f"PASS: Component 2 — Values preserved, D1 and columns A-C untouched (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Issues: {sub_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
