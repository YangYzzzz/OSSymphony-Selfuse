"""
Initial Setup: NPER formula for credit card debt payoff
Task ID: calc_fmb_nper_035
Domain: libreoffice_calc

Creates a spreadsheet with debt payoff parameters.
Cell B6 (Months to Payoff) is LEFT EMPTY — that is the target cell for the task.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_nper_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Debt Payoff'

    # Column widths for readability
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 18

    # Title row
    ws['A1'] = 'Credit Card Debt Payoff Calculator'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Data rows as specified in task context
    ws['A2'] = 'Current Balance'
    ws['B2'] = 8500

    ws['A3'] = 'Monthly Interest Rate'
    ws['B3'] = 0.015

    ws['A4'] = 'Monthly Payment'
    ws['B4'] = 250

    # Row 5 blank (as specified)
    ws['A5'] = ''
    ws['B5'] = ''

    # Row 6: label present, value EMPTY (this is the target cell)
    ws['A6'] = 'Months to Payoff'
    # B6 intentionally left empty — task is to put =NPER(B3,-B4,B2) here

    # Format B3 as percentage display
    ws['B3'].number_format = '0.0%'

    # Format B2 and B4 as currency
    ws['B2'].number_format = '$#,##0.00'
    ws['B4'].number_format = '$#,##0.00'

    # Bold the labels
    for row in range(2, 7):
        ws[f'A{row}'].font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: Debt Payoff')
    print('  A2=Current Balance, B2=8500')
    print('  A3=Monthly Interest Rate, B3=0.015')
    print('  A4=Monthly Payment, B4=250')
    print('  A5=blank, B5=blank')
    print('  A6=Months to Payoff, B6=EMPTY (task target)')


create_initial()
