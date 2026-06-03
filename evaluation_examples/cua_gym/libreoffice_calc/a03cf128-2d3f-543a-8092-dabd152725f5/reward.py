"""
Reward Script: Determine months to pay off credit card using NPER function
Task ID: calc_fmb_nper_035
Domain: libreoffice_calc
Scoring:
  - Component 1: B6 contains a formula (any formula present, not empty) — 0.3 points
  - Component 2: B6 formula uses NPER function — 0.3 points
  - Component 3: NPER formula has correct arguments referencing B3, -B4, B2 — 0.4 points
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_nper_035'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Put =NPER(B3,-B4,B2) in cell B6 of sheet 'Debt Payoff'.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Debt Payoff' sheet must exist
    if 'Debt Payoff' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Debt Payoff' not found in workbook. Sheets: {wb.sheetnames}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Debt Payoff']

    # Component 1: B6 contains any non-empty value or formula (0.3 points)
    # This verifies the agent placed something in the target cell
    try:
        b6_value = ws['B6'].value
        if b6_value is not None and str(b6_value).strip() != '':
            print(f"PASS: Component 1 — B6 is not empty (value: {repr(b6_value)}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — B6 is empty, expected an NPER formula")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B6 contains an NPER formula (0.3 points)
    # The task specifically requires using the NPER function
    try:
        b6_value = ws['B6'].value
        if b6_value is not None and isinstance(b6_value, str):
            formula_upper = b6_value.strip().upper().replace(' ', '')
            if formula_upper.startswith('=NPER('):
                print(f"PASS: Component 2 — B6 uses NPER function: {repr(b6_value)} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — B6 does not start with =NPER(, found: {repr(b6_value)}")
        else:
            print(f"FAIL: Component 2 — B6 is not a formula string, found: {repr(b6_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: NPER formula has correct arguments: B3, -B4, B2 (0.4 points)
    # Verifies the exact argument structure: rate=B3, nper_pmt=-B4, pv=B2
    try:
        b6_value = ws['B6'].value
        if b6_value is not None and isinstance(b6_value, str):
            # Normalize: strip spaces, uppercase
            formula_norm = b6_value.strip().upper().replace(' ', '')
            # Expected formula pattern: =NPER(B3,-B4,B2)
            # Allow for optional additional arguments (fv, type) with any value
            # Core requirement: rate=B3, pmt=-B4, pv=B2
            # Match =NPER(B3,-B4,B2) with optional ,... at end
            pattern = r'^=NPER\(B3,-B4,B2(\,.*)?[\)]?$'
            # Also check exact string match (most common form)
            exact_match = formula_norm == '=NPER(B3,-B4,B2)'
            regex_match = bool(re.match(pattern, formula_norm))

            if exact_match or regex_match:
                print(f"PASS: Component 3 — NPER formula has correct args (B3,-B4,B2): {repr(b6_value)} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 3 — NPER args incorrect. Expected =NPER(B3,-B4,B2), found: {repr(b6_value)}")
                print(f"  Normalized: {formula_norm}")
        else:
            print(f"FAIL: Component 3 — B6 is not a formula string: {repr(b6_value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
