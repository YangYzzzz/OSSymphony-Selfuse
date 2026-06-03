"""
Reward Script: Download COVID-19 CSV and filter for Germany in LibreOffice Calc
Task ID: osworld_multi_apps_sys_browser_os_005
Domain: libreoffice_calc (multi-app: browser + terminal + calc)
Scoring:
  Component 1 (0.3): CSV downloaded to /home/user/data/covid_data.csv with valid headers
  Component 2 (0.3): xlsx file exists with covid_data sheet containing the CSV data
  Component 3 (0.4): AutoFilter applied on Country column showing only Germany rows
                      (non-Germany rows hidden, filter value set to 'Germany')
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_sys_browser_os_005'

CSV_PATH = '/home/user/data/covid_data.csv'
XLSX_PATH = f'{WORKDIR}/{TASK_ID}.xlsx'


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Component 1: CSV file downloaded to /home/user/data/covid_data.csv (0.3 points)
    # This checks that the agent actually downloaded the file via wget/curl.
    # The initial_env has an empty /data/ folder — no csv present.
    try:
        if os.path.exists(CSV_PATH):
            with open(CSV_PATH, 'r') as f:
                first_line = f.readline().strip()
            if first_line and 'Country' in first_line and 'Confirmed' in first_line:
                print(f"PASS: Component 1 — CSV downloaded to {CSV_PATH} "
                      f"with valid headers: '{first_line}' (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 — CSV exists but has unexpected header: '{first_line}'")
        else:
            print(f"FAIL: Component 1 — CSV not found at {CSV_PATH}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: xlsx file exists with a sheet containing the COVID data (0.3 points)
    # Checks that LibreOffice Calc was opened with the CSV and the data is present.
    try:
        if not os.path.exists(XLSX_PATH):
            print(f"FAIL: Component 2 — xlsx file not found at {XLSX_PATH}")
        else:
            wb = openpyxl.load_workbook(XLSX_PATH)
            sheet_names = wb.sheetnames

            # Find a sheet that looks like the covid data
            target_sheet = None
            for name in sheet_names:
                if 'covid' in name.lower() or 'data' in name.lower():
                    target_sheet = name
                    break
            if target_sheet is None and len(sheet_names) > 0:
                target_sheet = sheet_names[0]

            if target_sheet:
                ws = wb[target_sheet]
                headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                has_country_col = any(
                    h and 'country' in str(h).lower() for h in headers
                )
                has_data_rows = ws.max_row > 5

                if has_country_col and has_data_rows:
                    print(f"PASS: Component 2 — xlsx has sheet '{target_sheet}' "
                          f"with headers {headers} and {ws.max_row} rows (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 2 — sheet '{target_sheet}' missing Country col "
                          f"or insufficient data. Headers: {headers}, rows: {ws.max_row}")
            else:
                print(f"FAIL: Component 2 — xlsx has no valid sheet. Sheets: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: AutoFilter set on Country column with Germany as filter value,
    # and non-Germany rows are hidden (0.4 points).
    # In initial_env the xlsx does not exist, so this always fails there.
    try:
        if not os.path.exists(XLSX_PATH):
            print(f"FAIL: Component 3 — xlsx file not found at {XLSX_PATH}")
        else:
            wb = openpyxl.load_workbook(XLSX_PATH)
            ws = wb.active

            auto_filter_ref = ws.auto_filter.ref
            if not auto_filter_ref:
                print(f"FAIL: Component 3 — No AutoFilter set on the sheet")
            else:
                # Check whether any filter column targets 'Germany'
                germany_filter_found = any(
                    fc.filters and fc.filters.filter and 'Germany' in [
                        str(v).strip() for v in fc.filters.filter
                    ]
                    for fc in ws.auto_filter.filterColumn
                )

                if not germany_filter_found:
                    print(f"FAIL: Component 3 — AutoFilter set ({auto_filter_ref}) but "
                          f"does not filter for 'Germany'")
                else:
                    # Verify that non-Germany rows are hidden
                    visible_germany = 0
                    hidden_non_germany = 0
                    visible_non_germany = 0

                    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                        row_num = row[0].row
                        country_val = row[1].value if len(row) > 1 else None
                        rd = ws.row_dimensions.get(row_num)
                        is_hidden = rd.hidden if rd else False

                        if country_val == 'Germany':
                            if not is_hidden:
                                visible_germany += 1
                        elif is_hidden:
                            hidden_non_germany += 1
                        else:
                            visible_non_germany += 1

                    all_germany_visible = visible_germany > 0
                    all_non_germany_hidden = visible_non_germany == 0 and hidden_non_germany > 0

                    if all_germany_visible and all_non_germany_hidden:
                        total_score += 0.4
                        print(f"PASS: Component 3 — Germany filter applied correctly: "
                              f"{visible_germany} Germany rows visible, "
                              f"{hidden_non_germany} non-Germany rows hidden (0.4 pts)")
                    elif germany_filter_found and hidden_non_germany > visible_non_germany:
                        total_score += 0.2
                        # Filter definition correct, majority of non-Germany rows hidden
                        print(f"PARTIAL: Component 3 — Germany filter set but incomplete. "
                              f"Germany visible: {visible_germany}, non-Germany hidden: {hidden_non_germany}, "
                              f"non-Germany still visible: {visible_non_germany} (0.2 pts)")
                    else:
                        print(f"FAIL: Component 3 — Filter not applied correctly. "
                              f"Germany visible: {visible_germany}, non-Germany hidden: {hidden_non_germany}, "
                              f"non-Germany still visible: {visible_non_germany}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task()
