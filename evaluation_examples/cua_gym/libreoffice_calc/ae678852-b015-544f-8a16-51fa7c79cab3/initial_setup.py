"""
Initial Setup: Add error bars to line chart showing standard deviation
Task ID: calc_chart_error_bars_069
Domain: libreoffice_calc

Creates a spreadsheet with an 'Experiment' sheet containing experimental data
(Condition, Mean Score, Std Dev) and a line chart plotting Mean Score WITHOUT
error bars. The task requires adding error bars using Std Dev values.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_error_bars_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = Workbook()

    # --- Sheet: Experiment ---
    ws = wb.active
    ws.title = 'Experiment'

    # Headers
    headers = ['Condition', 'Mean Score', 'Std Dev']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows (as specified in task context)
    data = [
        ['Control',     72.4, 8.2],
        ['Treatment A', 78.1, 6.5],
        ['Treatment B', 81.6, 7.1],
        ['Treatment C', 85.2, 5.8],
        ['Treatment D', 79.8, 9.4],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12

    # --- Line Chart (NO error bars) ---
    chart = LineChart()
    chart.title = 'Experimental Results'
    chart.style = 10
    chart.y_axis.title = 'Mean Score'
    chart.x_axis.title = 'Condition'
    chart.y_axis.crossAx = 500
    chart.x_axis.crossAx = 100

    # Data: Mean Score column (B), with header from row 1
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)

    # Categories: Condition column (A), rows 2-6
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.set_categories(cats)

    # Chart dimensions
    chart.width = 15
    chart.height = 10

    ws.add_chart(chart, 'E2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: Experiment')
    print('Data rows: 5 (Control, Treatment A-D)')
    print('Chart: Line chart with Mean Score — NO error bars')


create_initial()
