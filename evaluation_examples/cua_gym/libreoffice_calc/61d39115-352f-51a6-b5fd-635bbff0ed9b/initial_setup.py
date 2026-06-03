"""
Initial Setup: AI Ethics Researchers Directory - Pre-task state
Task ID: osworld_multi_apps_web_prof_email_012
Domain: libreoffice_calc

Creates AI_Ethics_Researchers.xlsx with Sheet1 containing researcher data
(Name, University, Homepage, Email, Title) where Email and Title are blank.
Opens the file in LibreOffice Calc and also opens Chrome.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_012'
OUTPUT = f'{WORKDIR}/AI_Ethics_Researchers.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Researcher Directory ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers
    headers = ['Name', 'University', 'Homepage', 'Email', 'Title']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Researcher data — realistic AI ethics professors from real universities
    # Email and Title are intentionally left blank (task requires agent to fill them)
    researchers = [
        {
            'name': 'Timnit Gebru',
            'university': 'Stanford University',
            'homepage': 'https://ai.stanford.edu/~tgebru/',
            'email': '',
            'title': '',
        },
        {
            'name': 'Kate Crawford',
            'university': 'New York University',
            'homepage': 'https://ainowinstitute.org/people/kate-crawford.html',
            'email': '',
            'title': '',
        },
        {
            'name': 'Ruha Benjamin',
            'university': 'Princeton University',
            'homepage': 'https://www.ruhabenjamin.com',
            'email': '',
            'title': '',
        },
        {
            'name': 'Safiya Umoja Noble',
            'university': 'University of California Los Angeles',
            'homepage': 'https://safiyaunoble.com',
            'email': '',
            'title': '',
        },
        {
            'name': 'Bettina Berendt',
            'university': 'KU Leuven',
            'homepage': 'https://people.cs.kuleuven.be/~bettina.berendt/',
            'email': '',
            'title': '',
        },
    ]

    for row_idx, researcher in enumerate(researchers, 2):
        ws1.cell(row=row_idx, column=1, value=researcher['name'])
        ws1.cell(row=row_idx, column=2, value=researcher['university'])
        ws1.cell(row=row_idx, column=3, value=researcher['homepage'])
        ws1.cell(row=row_idx, column=4, value=researcher['email'])   # blank
        ws1.cell(row=row_idx, column=5, value=researcher['title'])   # blank

    # Column widths for readability
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 35
    ws1.column_dimensions['C'].width = 50
    ws1.column_dimensions['D'].width = 30
    ws1.column_dimensions['E'].width = 35

    # Freeze header row
    ws1.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup — open Chrome first, then LibreOffice Calc
    launch_gui('google-chrome --new-window', delay_sec=2.0)
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
