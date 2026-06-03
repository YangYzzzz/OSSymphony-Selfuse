"""
Reward Script: Compile AI Ethics Researcher Directory
Task ID: osworld_multi_apps_web_prof_email_012
Domain: libreoffice_calc
Scoring:
  - Component 1: 'Publications' sheet exists with correct columns              (0.2 pts)
  - Component 2: Publications sheet has 15 data rows (3 per researcher)        (0.3 pts)
  - Component 3: Sheet1 Email column filled for all 5 researchers              (0.25 pts)
  - Component 4: Sheet1 Title column filled for all 5 researchers              (0.25 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_012'
FILE_NAME = 'AI_Ethics_Researchers.xlsx'
FILE_PATH = os.path.join(WORKDIR, FILE_NAME)

EXPECTED_RESEARCHERS = [
    'Timnit Gebru',
    'Kate Crawford',
    'Ruha Benjamin',
    'Safiya Umoja Noble',
    'Bettina Berendt',
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: 'Publications' sheet exists with correct columns (0.2 pts)
    # This FAILS on initial (no Publications sheet) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        if 'Publications' in wb.sheetnames:
            ws_pub = wb['Publications']
            header = [ws_pub.cell(row=1, column=c).value for c in range(1, 4)]
            expected_headers = ['Researcher Name', 'Publication Title', 'Year']
            headers_match = all(
                h is not None and str(h).strip().lower() == exp.lower()
                for h, exp in zip(header, expected_headers)
            )
            if headers_match:
                print(f"PASS: Component 1 — 'Publications' sheet exists with correct headers {header} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 1 — 'Publications' sheet found but headers incorrect: {header}")
        else:
            print(f"FAIL: Component 1 — 'Publications' sheet not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Publications sheet has 15 data rows (3 per researcher) (0.3 pts)
    # This FAILS on initial (no Publications sheet) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        if 'Publications' not in wb.sheetnames:
            print(f"FAIL: Component 2 — 'Publications' sheet not found")
        else:
            ws_pub = wb['Publications']
            # Collect valid data rows (skip header row=1)
            data_rows = []
            for row in ws_pub.iter_rows(min_row=2):
                row_vals = [c.value for c in row]
                if row_vals[0] is not None and row_vals[1] is not None:
                    data_rows.append(row_vals)

            num_data_rows = len(data_rows)

            # Count publications per researcher
            researcher_counts = {}
            for row in data_rows:
                name = str(row[0]).strip()
                researcher_counts[name] = researcher_counts.get(name, 0) + 1

            researchers_with_three = sum(
                1 for name in EXPECTED_RESEARCHERS if researcher_counts.get(name, 0) >= 3
            )
            component2_score = (researchers_with_three / len(EXPECTED_RESEARCHERS)) * 0.3

            if researchers_with_three == len(EXPECTED_RESEARCHERS) and num_data_rows >= 15:
                print(f"PASS: Component 2 — {num_data_rows} publication rows, all researchers have >= 3 entries (0.3 pts)")
                print(f"  Researcher publication counts: {researcher_counts}")
                total_score += 0.3
            elif component2_score > 0:
                print(f"FAIL: Component 2 — {researchers_with_three}/5 researchers have >= 3 publications. Partial: {component2_score:.3f} pts")
                print(f"  Researcher publication counts: {researcher_counts}")
                if component2_score > 0:
                    total_score += component2_score
            else:
                print(f"FAIL: Component 2 — No valid publication data rows found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Sheet1 Email column filled for all 5 researchers (0.25 pts)
    # This FAILS on initial (Email is None for all) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        ws1 = wb['Sheet1']
        header_row = [ws1.cell(row=1, column=c).value for c in range(1, ws1.max_column + 1)]
        if 'Email' not in header_row:
            print(f"FAIL: Component 3 — 'Email' column not found in Sheet1. Headers: {header_row}")
        else:
            email_col = header_row.index('Email') + 1  # 1-indexed
            emails_filled = 0
            for r in range(2, ws1.max_row + 1):
                email_val = ws1.cell(row=r, column=email_col).value
                name_val = ws1.cell(row=r, column=1).value
                if name_val is not None and email_val is not None and str(email_val).strip() != '':
                    emails_filled += 1
                    print(f"  Email found: {name_val} -> {email_val}")
                elif name_val is not None:
                    print(f"  Email MISSING for: {name_val}")

            component3_score = (emails_filled / 5.0) * 0.25
            if emails_filled >= 5:
                print(f"PASS: Component 3 — All {emails_filled} researchers have email filled (0.25 pts)")
                total_score += 0.25
            elif component3_score > 0:
                print(f"FAIL: Component 3 — Only {emails_filled}/5 researchers have email. Partial: {component3_score:.3f} pts")
                if component3_score > 0:
                    total_score += component3_score
            else:
                print(f"FAIL: Component 3 — No emails filled in Sheet1")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Sheet1 Title column filled for all 5 researchers (0.25 pts)
    # This FAILS on initial (Title is None for all) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        ws1 = wb['Sheet1']
        header_row = [ws1.cell(row=1, column=c).value for c in range(1, ws1.max_column + 1)]
        if 'Title' not in header_row:
            print(f"FAIL: Component 4 — 'Title' column not found in Sheet1. Headers: {header_row}")
        else:
            title_col = header_row.index('Title') + 1  # 1-indexed
            titles_filled = 0
            for r in range(2, ws1.max_row + 1):
                title_val = ws1.cell(row=r, column=title_col).value
                name_val = ws1.cell(row=r, column=1).value
                if name_val is not None and title_val is not None and str(title_val).strip() != '':
                    titles_filled += 1
                    print(f"  Title found: {name_val} -> {title_val}")
                elif name_val is not None:
                    print(f"  Title MISSING for: {name_val}")

            component4_score = (titles_filled / 5.0) * 0.25
            if titles_filled >= 5:
                print(f"PASS: Component 4 — All {titles_filled} researchers have title filled (0.25 pts)")
                total_score += 0.25
            elif component4_score > 0:
                print(f"FAIL: Component 4 — Only {titles_filled}/5 researchers have title. Partial: {component4_score:.3f} pts")
                if component4_score > 0:
                    total_score += component4_score
            else:
                print(f"FAIL: Component 4 — No titles filled in Sheet1")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Entrypoint: test against canonical artifact path on this VM
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
