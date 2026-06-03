"""
Initial Setup: Create workbook with Q1, Q2 data sheets and empty Summary sheet
Task ID: calc_chart_multi_sheet_series_068
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_multi_sheet_series_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Q1 ---
    ws_q1 = wb.active
    ws_q1.title = 'Q1'

    # Headers
    ws_q1['A1'] = 'Month'
    ws_q1['B1'] = 'Revenue'

    # Q1 data: January, February, March
    q1_data = [
        ('January', 420000),
        ('February', 385000),
        ('March', 460000),
    ]
    for r, (month, revenue) in enumerate(q1_data, 2):
        ws_q1.cell(row=r, column=1, value=month)
        ws_q1.cell(row=r, column=2, value=revenue)

    # --- Sheet 2: Q2 ---
    ws_q2 = wb.create_sheet('Q2')

    # Headers
    ws_q2['A1'] = 'Month'
    ws_q2['B1'] = 'Revenue'

    # Q2 data: April, May, June
    q2_data = [
        ('April', 495000),
        ('May', 528000),
        ('June', 572000),
    ]
    for r, (month, revenue) in enumerate(q2_data, 2):
        ws_q2.cell(row=r, column=1, value=month)
        ws_q2.cell(row=r, column=2, value=revenue)

    # --- Sheet 3: Summary (empty — no charts, no data) ---
    ws_summary = wb.create_sheet('Summary')
    # Leave completely empty — the task is to create a chart here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Q1 (3 months data), Q2 (3 months data), Summary (empty)')


create_initial()
