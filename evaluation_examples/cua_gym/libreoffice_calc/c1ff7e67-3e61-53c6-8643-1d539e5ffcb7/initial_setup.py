"""
Initial Setup: Apply conditional formatting for duplicate respondent IDs
Task ID: calc_gg3_041
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg3_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Submissions'

    # Header row
    ws.cell(row=1, column=1, value='Respondent ID')
    ws['A1'].font = Font(bold=True, size=11, name='Calibri')
    ws['A1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['A1'].font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.cell(row=1, column=2, value='Submission Date')
    ws['B1'].font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
    ws['B1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['B1'].alignment = Alignment(horizontal='center')

    ws.cell(row=1, column=3, value='Survey Score')
    ws['C1'].font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')
    ws['C1'].fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    ws['C1'].alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14

    # Generate respondent IDs with deliberate duplicates
    # We'll create a pool of unique IDs, then introduce duplicates
    random.seed(42)  # reproducible

    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
        'Carlos', 'Fatima', 'Robert', 'Aisha', 'Thomas', 'Yuki', 'Michael',
        'Zara', 'Daniel', 'Olivia', 'Ahmed', 'Sofia', 'Nathan', 'Grace',
        'Leo', 'Maya', 'Ryan', 'Hannah', 'Victor', 'Nina', 'Samuel',
        'Lina', 'Kevin', 'Julia', 'Omar', 'Chloe', 'Ivan', 'Emma',
        'Pedro', 'Anna', 'Felix', 'Diana', 'Hugo', 'Rachel', 'Anton',
        'Leila', 'Chris', 'Tara', 'Nils', 'Beth', 'Ravi', 'Mona', 'Sean',
        'Iris', 'Joel', 'Vera', 'Kurt', 'Sana', 'Axel', 'Dina', 'Phil',
        'Rosa', 'Evan', 'Lily', 'Max', 'Noor', 'Brent', 'Ada', 'Tom',
        'Ines', 'Dale', 'Sue', 'Roy'
    ]
    last_names = [
        'Chen', 'Johnson', 'Patel', 'Williams', 'Kumar', 'Brown', 'Liu',
        'Garcia', 'Ali', 'Davis', 'Hassan', 'Wilson', 'Tanaka', 'Taylor',
        'Khan', 'Moore', 'Martin', 'Ibrahim', 'Rodriguez', 'Clark',
        'Thompson', 'Kim', 'Gupta', 'Anderson', 'Miller', 'Petrov', 'Reyes',
        'Walker', 'Muller', 'Harris', 'Scott', 'Rivera', 'Lee', 'Young',
        'Adams', 'Baker', 'Gonzalez', 'Nelson', 'Carter', 'Mitchell',
        'Roberts', 'Turner', 'Phillips', 'Campbell', 'Parker', 'Evans',
        'Edwards', 'Collins', 'Stewart', 'Sanchez'
    ]

    # Generate 70 unique IDs (format: RESP-XXXX)
    unique_ids = []
    used_nums = set()
    for i in range(70):
        while True:
            num = random.randint(1000, 9999)
            if num not in used_nums:
                used_nums.add(num)
                unique_ids.append(f'RESP-{num}')
                break

    # Build 100-row data: 70 unique + 30 duplicates of some existing ones
    respondent_ids = list(unique_ids)
    # Pick 30 IDs to duplicate (repeat from the first 50)
    duplicate_sources = random.sample(unique_ids[:50], 30)
    respondent_ids.extend(duplicate_sources)

    # Shuffle to mix duplicates among originals
    random.shuffle(respondent_ids)
    respondent_ids = respondent_ids[:100]

    # Dates for submissions (2025-01 to 2025-03)
    months = ['2025-01', '2025-02', '2025-03']

    for i, resp_id in enumerate(respondent_ids):
        row = i + 2
        ws.cell(row=row, column=1, value=resp_id)
        month = random.choice(months)
        day = random.randint(1, 28)
        ws.cell(row=row, column=2, value=f'{month}-{day:02d}')
        ws.cell(row=row, column=3, value=random.randint(35, 100))

    # Freeze header row
    ws.freeze_panes = 'A2'

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
