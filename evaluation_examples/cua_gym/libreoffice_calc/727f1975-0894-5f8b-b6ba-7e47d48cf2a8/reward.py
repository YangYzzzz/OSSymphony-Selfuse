"""
FINAL REWARD SCRIPT - SUCCESS
Task: Please copy the values from 'Part Number' to 'Standardized Part' and add leading zeros to make them all 9 digits long. Don't modify any other cells.
Generated: 2025-11-24 07:27:03
Status: success
Model: o3
Total Steps: 6
"""

import openpyxl
import math
import os
import pathlib
import re
import sys


def locate_workbook(path_argument: str | None = None) -> pathlib.Path | None:
    """Locate the workbook to verify.

    Search order:
    1. Explicit path supplied as a CLI argument.
    2. Environment variable TASK_FILE (if set by evaluator).
    3. Any *.xlsx file in current directory whose name contains the word "golden".
    4. First *.xlsx file in current directory.
    """
    if path_argument and pathlib.Path(path_argument).is_file():
        return pathlib.Path(path_argument)

    env_path = os.getenv("TASK_FILE")
    if env_path and pathlib.Path(env_path).is_file():
        return pathlib.Path(env_path)

    cwd = pathlib.Path.cwd()
    golden = sorted(cwd.glob("*golden*.xlsx"))
    if golden:
        return golden[0]

    generic = sorted(cwd.glob("*.xlsx"))
    if generic:
        return generic[0]

    return None


def verify_task(file_path: pathlib.Path) -> float:
    """Verify that the values from column 'Part Number' have been copied to
    'Standardized Part' and padded to 9-digit strings, with no other cells changed.
    A progressive score (0.0-1.0) is returned based on row-level accuracy.
    """
    MAX_SCORE = 1.0

    print(f"Verifying Excel task for file: {file_path}\n")

    # ---------------------------------------------------------
    # 1️⃣  Load workbook (data_only so we read stored values)
    # ---------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ---------------------------------------------------------
    # 2️⃣  Locate expected sheet
    # ---------------------------------------------------------
    sheet_name = "Parts List"
    if sheet_name not in wb.sheetnames:
        print(f"✗ Expected sheet '{sheet_name}' not found. Sheets present: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0
    ws = wb[sheet_name]

    # ---------------------------------------------------------
    # 3️⃣  Validate header integrity (task forbids other edits)
    # ---------------------------------------------------------
    header_values = [cell.value for cell in ws[1]]
    print(f"Header row values: {header_values}")

    expected_header = ["Part Number", "Standardized Part", "Description"]
    if header_values[:3] != expected_header:
        print("✗ Header row has been altered – other cells must remain unchanged.")
        print("REWARD: 0.0")
        return 0.0

    # Column indices (1-based for openpyxl)
    col_part = header_values.index("Part Number") + 1
    col_std = header_values.index("Standardized Part") + 1

    # ---------------------------------------------------------
    # 4️⃣  Row-by-row verification
    # ---------------------------------------------------------
    total_rows = 0  # data rows seen (skip completely blank tails)
    correct_rows = 0

    for row_idx in range(2, ws.max_row + 1):
        part_cell = ws.cell(row=row_idx, column=col_part)
        std_cell = ws.cell(row=row_idx, column=col_std)
        part_val = part_cell.value
        std_val = std_cell.value

        # Ignore trailing completely empty rows
        if part_val is None and std_val is None:
            continue

        total_rows += 1

        # ---------- 4a. Build expected standardized value ----------
        if part_val is None:
            print(f"Row {row_idx}: Missing Part Number – cannot standardize.")
            continue

        # Coerce numeric part value to int if it's a whole number (avoid 123.0)
        if isinstance(part_val, float) and math.isclose(part_val, int(part_val)):
            part_val = int(part_val)
        part_str = str(part_val).strip()

        # Part number must be purely digits
        if not re.fullmatch(r"\d+", part_str):
            print(f"Row {row_idx}: Part Number '{part_str}' is not numeric – invalid.")
            continue

        expected_std = part_str.zfill(9)

        # ---------- 4b. Extract actual standardized string ----------
        if std_val is None:
            actual_str = ""
        elif isinstance(std_val, (int, float)):
            # Convert numeric to int (if whole) then str (leading zeros lost). We'll check cell format.
            if isinstance(std_val, float) and math.isclose(std_val, int(std_val)):
                std_val = int(std_val)
            actual_str = str(int(std_val))
        else:
            actual_str = str(std_val).strip()

        # ---------- 4c. Validation logic ----------
        row_correct = False

        # Case 1 – stored literally as the correct 9-digit string
        if isinstance(std_val, str) and actual_str == expected_std:
            row_correct = True

        # Case 2 – stored numerically but formatted with 9-digit mask ("000000000")
        elif isinstance(std_val, (int, float)):
            numeric_match = int(float(std_val)) == int(part_str)
            format_match = std_cell.number_format == "000000000"
            if numeric_match and format_match:
                row_correct = True

        if row_correct:
            correct_rows += 1
        else:
            print(
                f"Row {row_idx}: Standardized Part '{actual_str}' incorrect – expected '{expected_std}'."
            )

    # ---------------------------------------------------------
    # 5️⃣  Scoring
    # ---------------------------------------------------------
    if total_rows == 0:
        print("✗ No data rows found – nothing to evaluate.")
        print("REWARD: 0.0")
        return 0.0

    accuracy = correct_rows / total_rows
    print(f"\nSummary: {correct_rows} / {total_rows} rows correct.")
    print(f"Accuracy: {accuracy:.2%}")

    reward = round(min(MAX_SCORE, accuracy * MAX_SCORE), 2)
    print(f"REWARD: {reward}")
    return reward


# =============================================================
#                     Script Entry Point
# =============================================================
if __name__ == "__main__":
    arg_path = sys.argv[1] if len(sys.argv) > 1 else None
    workbook_path = locate_workbook(arg_path)

    if not workbook_path:
        print("✗ No .xlsx file found to verify in current directory.")
        print("REWARD: 0.0")
        sys.exit(0)

    verify_task(workbook_path)

