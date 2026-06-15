"""
Reward Script: Fill Email and Affiliation for conference invitees from faculty webpages
Task ID: osworld_multi_apps_web_prof_email_005
Domain: libreoffice_calc
Scoring:
  - Component 1: Email column fully populated (all 7 rows, valid email format) — 0.5 pts
  - Component 2: Affiliation column fully populated (all 7 rows, non-empty strings) — 0.5 pts
  Total: 1.0
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_005'
FILE_PATH = os.path.join(WORKDIR, 'Conference_Invitees.xlsx')

# Professors in the spreadsheet (rows 2-8, 7 total)
EXPECTED_ROW_COUNT = 7

# Simple email pattern (not exhaustive, but catches obviously malformed values)
EMAIL_PATTERN = re.compile(r'^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$')

# University keywords we expect to see in affiliations
UNIVERSITY_KEYWORDS = [
    'university', 'institute', 'school', 'department', 'lab', 'college',
    'berkeley', 'mit', 'stanford', 'cmu', 'carnegie', 'cornell',
]


def is_valid_email(value):
    """Return True if value looks like a valid email address."""
    if not value or not isinstance(value, str):
        return False
    return bool(EMAIL_PATTERN.match(value.strip()))


def is_valid_affiliation(value):
    """Return True if value is a non-empty string containing a university/institution keyword."""
    if not value or not isinstance(value, str):
        return False
    lower = value.strip().lower()
    if len(lower) < 3:
        return False
    return any(kw in lower for kw in UNIVERSITY_KEYWORDS)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook — if this fails, the file is broken and we return 0.0
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate active sheet
    try:
        ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find Email and Affiliation column indices from header row (row 1)
    try:
        headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
        email_col = None
        affil_col = None
        for idx, h in enumerate(headers, start=1):
            if h and str(h).strip().lower() == 'email':
                email_col = idx
            if h and str(h).strip().lower() == 'affiliation':
                affil_col = idx
        if email_col is None or affil_col is None:
            print(f"CRITICAL: Could not find Email/Affiliation columns. Headers: {headers}")
            print("REWARD: 0.0")
            return 0.0
        print(f"INFO: Headers found — Email col={email_col}, Affiliation col={affil_col}")
    except Exception as e:
        print(f"CRITICAL: Error reading header row: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Determine data row range
    # Row 1 = header; data starts at row 2
    data_rows = list(range(2, ws.max_row + 1))
    row_count = len(data_rows)
    print(f"INFO: Data rows detected: {row_count} (expected {EXPECTED_ROW_COUNT})")

    if row_count == 0:
        print("FAIL: No data rows found in spreadsheet.")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # -----------------------------------------------------------------------
    # Component 1: Email column is fully populated with valid-looking emails
    #              (0.5 points)
    # This FAILS on initial_env (all empty) and PASSES on golden_env (all filled)
    # -----------------------------------------------------------------------
    try:
        email_values = []
        for r in data_rows:
            val = ws.cell(row=r, column=email_col).value
            email_values.append(val)

        valid_emails = [v for v in email_values if is_valid_email(str(v) if v is not None else '')]
        total_data = len(data_rows)
        valid_count = len(valid_emails)

        if valid_count == total_data:
            print(f"PASS: Component 1 — All {total_data} Email cells populated with valid emails (0.5 pts)")
            print(f"      Emails: {email_values}")
            total_score += 0.5
        elif valid_count > 0:
            partial = round(0.5 * valid_count / total_data, 4)
            print(f"FAIL: Component 1 — Only {valid_count}/{total_data} Email cells populated; awarding partial 0.0 pts")
            print(f"      Emails: {email_values}")
            # Partial credit within component: award nothing unless at least all 7 are valid
            # (strict: all-or-nothing for the component to keep clean 0.0 on initial_env)
        else:
            print(f"FAIL: Component 1 — Email column empty ({valid_count}/{total_data} valid)")
            print(f"      Emails: {email_values}")
    except Exception as e:
        print(f"ERROR: Component 1 (Email check) — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Affiliation column is fully populated with institution names
    #              (0.5 points)
    # This FAILS on initial_env (all empty) and PASSES on golden_env (all filled)
    # -----------------------------------------------------------------------
    try:
        affil_values = []
        for r in data_rows:
            val = ws.cell(row=r, column=affil_col).value
            affil_values.append(val)

        valid_affils = [v for v in affil_values if is_valid_affiliation(str(v) if v is not None else '')]
        valid_affil_count = len(valid_affils)

        if valid_affil_count == total_data:
            print(f"PASS: Component 2 — All {total_data} Affiliation cells populated with institution names (0.5 pts)")
            print(f"      Affiliations: {affil_values}")
            total_score += 0.5
        elif valid_affil_count > 0:
            print(f"FAIL: Component 2 — Only {valid_affil_count}/{total_data} Affiliation cells populated with institution names")
            print(f"      Affiliations: {affil_values}")
        else:
            print(f"FAIL: Component 2 — Affiliation column empty ({valid_affil_count}/{total_data} valid)")
            print(f"      Affiliations: {affil_values}")
    except Exception as e:
        print(f"ERROR: Component 2 (Affiliation check) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
