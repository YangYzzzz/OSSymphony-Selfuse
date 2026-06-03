"""
Initial Setup: Conference Invitees spreadsheet with professor info for web lookup
Task ID: osworld_multi_apps_web_prof_email_005
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_multi_apps_web_prof_email_005'
OUTPUT = f'{WORKDIR}/Conference_Invitees.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conference Invitees"

    # --- Column headers ---
    headers = ['Name', 'Webpage', 'Email', 'Affiliation']
    header_font = Font(name='Calibri', size=12, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_color = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=12, bold=True, color='FFFFFFFF')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = header_align
        cell.border = header_border

    # --- Professor data (Name, Webpage — Email and Affiliation are intentionally BLANK) ---
    # These are fictional professors at well-known universities with plausible academic URLs
    professors = [
        ['Dr. Elena Marchetti',
         'https://www2.eecs.berkeley.edu/Faculty/Homepages/marchetti.html',
         '',   # Email — BLANK (to be filled by agent)
         ''],  # Affiliation — BLANK (to be filled by agent)
        ['Prof. James R. Sutherland',
         'https://www.csail.mit.edu/person/james-sutherland',
         '',
         ''],
        ['Dr. Priya Nair',
         'https://cs.stanford.edu/people/priya-nair',
         '',
         ''],
        ['Prof. Kevin L. Brennan',
         'https://www.cs.cmu.edu/~kbrennan',
         '',
         ''],
        ['Dr. Sophia T. Walton',
         'https://www.cs.cornell.edu/people/sophia-walton',
         '',
         ''],
        ['Prof. Daniel Osei',
         'https://www2.eecs.berkeley.edu/Faculty/Homepages/osei.html',
         '',
         ''],
        ['Dr. Laura Kim',
         'https://www.csail.mit.edu/person/laura-kim',
         '',
         ''],
    ]

    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, row_data in enumerate(professors, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = data_align
            cell.border = data_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 28   # Name
    ws.column_dimensions['B'].width = 60   # Webpage
    ws.column_dimensions['C'].width = 38   # Email
    ws.column_dimensions['D'].width = 35   # Affiliation

    # --- Row heights ---
    ws.row_dimensions[1].height = 22  # Header row
    for r in range(2, len(professors) + 2):
        ws.row_dimensions[r].height = 20

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # --- GUI-ready startup ---
    # Open Chrome first (agent needs to browse professor pages)
    launch_gui('google-chrome', delay_sec=3.0)
    # Open the spreadsheet in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
