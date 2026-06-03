"""
Reward Script: Sales order validation with data validations and input messages
Task ID: calc_nrv_090
Domain: libreoffice_calc
Scoring: 6 components (one per validated cell B2-G2), each checking validation type,
         parameters, and presence of input message. Total = 1.0.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_090'


def find_validation_for_cell(ws, cell_ref):
    """Find the DataValidation object that applies to a given cell reference."""
    from openpyxl.utils import coordinate_to_tuple
    row, col = coordinate_to_tuple(cell_ref)
    for dv in ws.data_validations.dataValidation:
        # dv.sqref is a CellRange or MultiCellRange; check if cell_ref is in it
        if cell_ref in str(dv.sqref):
            return dv
        # Also check by iterating cells in the range
        try:
            for rng in dv.sqref.ranges:
                if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                    return dv
        except AttributeError:
            # sqref might be a single CellRange, not MultiCellRange
            try:
                if dv.sqref.min_row <= row <= dv.sqref.max_row and dv.sqref.min_col <= col <= dv.sqref.max_col:
                    return dv
            except AttributeError:
                pass
    return None


def has_input_message(dv):
    """Check if a DataValidation has an input message configured."""
    return dv.prompt is not None and len(str(dv.prompt).strip()) > 0


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the OrderForm sheet (or active sheet)
    ws = None
    if 'OrderForm' in wb.sheetnames:
        ws = wb['OrderForm']
    else:
        ws = wb.active
    print(f"Checking sheet: {ws.title}")

    num_validations = len(ws.data_validations.dataValidation)
    print(f"Total data validations found: {num_validations}")

    # Component 1: B2 — textLength validation, operator=equal, formula1=6, with input message (0.17 pts)
    try:
        dv = find_validation_for_cell(ws, 'B2')
        if dv is not None and dv.type == 'textLength' and dv.operator == 'equal' and str(dv.formula1) == '6' and has_input_message(dv):
            print(f"PASS: Component 1 — B2 textLength=6 with input message (0.17 pts)")
            total_score += 0.17
        elif dv is not None and dv.type == 'textLength':
            # Partial: correct type but wrong params or missing message
            detail = f"type={dv.type}, op={dv.operator}, f1={dv.formula1}, msg={dv.prompt is not None}"
            print(f"FAIL: Component 1 — B2 textLength found but params incorrect: {detail}")
        else:
            found_type = dv.type if dv else 'None'
            print(f"FAIL: Component 1 — B2 expected textLength validation, found: {found_type}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C2 — date validation, operator>=TODAY(), with input message (0.17 pts)
    try:
        dv = find_validation_for_cell(ws, 'C2')
        if dv is not None and dv.type == 'date' and dv.operator == 'greaterThanOrEqual' and 'TODAY()' in str(dv.formula1).upper() and has_input_message(dv):
            print(f"PASS: Component 2 — C2 date>=TODAY() with input message (0.17 pts)")
            total_score += 0.17
        elif dv is not None and dv.type == 'date':
            detail = f"op={dv.operator}, f1={dv.formula1}, msg={dv.prompt is not None}"
            print(f"FAIL: Component 2 — C2 date found but params incorrect: {detail}")
        else:
            found_type = dv.type if dv else 'None'
            print(f"FAIL: Component 2 — C2 expected date validation, found: {found_type}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: D2 — list validation from ProductList named range, with input message (0.17 pts)
    try:
        dv = find_validation_for_cell(ws, 'D2')
        if dv is not None and dv.type == 'list' and 'ProductList' in str(dv.formula1) and has_input_message(dv):
            print(f"PASS: Component 3 — D2 list from ProductList with input message (0.17 pts)")
            total_score += 0.17
        elif dv is not None and dv.type == 'list':
            detail = f"formula1={dv.formula1}, msg={dv.prompt is not None}"
            print(f"FAIL: Component 3 — D2 list found but params incorrect: {detail}")
        else:
            found_type = dv.type if dv else 'None'
            print(f"FAIL: Component 3 — D2 expected list validation, found: {found_type}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: E2 — whole number 1-9999, with input message (0.16 pts)
    try:
        dv = find_validation_for_cell(ws, 'E2')
        if dv is not None and dv.type == 'whole' and dv.operator == 'between' and str(dv.formula1) == '1' and str(dv.formula2) == '9999' and has_input_message(dv):
            print(f"PASS: Component 4 — E2 whole 1-9999 with input message (0.16 pts)")
            total_score += 0.16
        elif dv is not None and dv.type == 'whole':
            detail = f"op={dv.operator}, f1={dv.formula1}, f2={dv.formula2}, msg={dv.prompt is not None}"
            print(f"FAIL: Component 4 — E2 whole found but params incorrect: {detail}")
        else:
            found_type = dv.type if dv else 'None'
            print(f"FAIL: Component 4 — E2 expected whole validation, found: {found_type}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: F2 — decimal 0-0.5, with input message (0.16 pts)
    try:
        dv = find_validation_for_cell(ws, 'F2')
        if dv is not None and dv.type == 'decimal' and dv.operator == 'between':
            f1_ok = abs(float(dv.formula1) - 0.0) < 0.001
            f2_ok = abs(float(dv.formula2) - 0.5) < 0.001
            if f1_ok and f2_ok and has_input_message(dv):
                print(f"PASS: Component 5 — F2 decimal 0-0.5 with input message (0.16 pts)")
                total_score += 0.16
            else:
                print(f"FAIL: Component 5 — F2 decimal range or message incorrect: f1={dv.formula1}, f2={dv.formula2}, msg={dv.prompt is not None}")
        elif dv is not None:
            print(f"FAIL: Component 5 — F2 expected decimal validation, found: type={dv.type}")
        else:
            print(f"FAIL: Component 5 — F2 no validation found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: G2 — list with 4 payment methods (Credit Card, Bank Transfer, Check, Cash), with input message (0.17 pts)
    try:
        dv = find_validation_for_cell(ws, 'G2')
        if dv is not None and dv.type == 'list' and has_input_message(dv):
            formula_str = str(dv.formula1).lower()
            # Check all 4 payment methods are present in the formula
            required = ['credit card', 'bank transfer', 'check', 'cash']
            all_present = all(method in formula_str for method in required)
            if all_present:
                print(f"PASS: Component 6 — G2 list with 4 payment methods and input message (0.17 pts)")
                total_score += 0.17
            else:
                missing = [m for m in required if m not in formula_str]
                print(f"FAIL: Component 6 — G2 list missing methods: {missing}. formula1={dv.formula1}")
        elif dv is not None:
            print(f"FAIL: Component 6 — G2 expected list validation, found: type={dv.type}")
        else:
            print(f"FAIL: Component 6 — G2 no validation found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
