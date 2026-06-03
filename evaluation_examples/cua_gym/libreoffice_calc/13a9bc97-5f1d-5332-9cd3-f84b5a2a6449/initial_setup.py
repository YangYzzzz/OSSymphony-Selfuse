"""
Initial Setup: Sales order validation form with headers and product list
Task ID: calc_nrv_090
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: OrderForm ---
    ws = wb.active
    ws.title = 'OrderForm'

    # Headers in row 1
    headers = ['Order #', 'Customer ID', 'Order Date', 'Product',
               'Quantity', 'Discount %', 'Payment Method']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Set column widths for readability
    col_widths = {'A': 12, 'B': 16, 'C': 14, 'D': 22, 'E': 12, 'F': 14, 'G': 20}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # A2 has a sample order number; B2:G2 are empty (agent fills them)
    ws.cell(row=1, column=1).value = 'Order #'
    ws.cell(row=2, column=1, value='ORD-2026-0001')
    ws.cell(row=2, column=1).font = Font(name='Calibri', size=11)
    ws.cell(row=2, column=1).border = thin_border

    # Light border on B2:G2 so the form looks prepared
    for col in range(2, 8):
        cell = ws.cell(row=2, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add a few more blank form rows (3-6) with borders for future orders
    for row in range(3, 7):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = thin_border

    # --- Sheet 2: Products ---
    ws2 = wb.create_sheet('Products')
    ws2.cell(row=1, column=1, value='Product Name')
    ws2.cell(row=1, column=1).font = Font(bold=True)
    ws2.column_dimensions['A'].width = 30

    products = [
        'Laptop Pro 15"', 'Desktop Workstation X7', 'Wireless Mouse M300',
        'Mechanical Keyboard K950', 'USB-C Hub 7-Port', '27" 4K Monitor',
        'Webcam HD 1080p', 'Noise-Cancelling Headset', 'Portable SSD 1TB',
        'External GPU Dock', 'Thunderbolt Cable 2m', 'Monitor Stand Adjustable',
        'Ergonomic Chair Pro', 'Standing Desk Converter', 'Cable Management Kit',
        'Docking Station Ultra', 'Wireless Charger Pad', 'Bluetooth Speaker S5',
        'Laptop Sleeve 15"', 'Screen Protector Matte', 'RAM Module 16GB DDR5',
        'NVMe SSD 2TB', 'Graphics Card RTX 4070', 'CPU Cooler AIO 240mm',
        'Power Supply 850W', 'PC Case Mid-Tower', 'Ethernet Cable Cat6 5m',
        'WiFi 6E Adapter', 'Surge Protector 8-Outlet', 'UPS Battery Backup 1500VA',
        'Thermal Paste Premium', 'Anti-Static Wrist Strap', 'Compressed Air Duster',
        'Cable Ties 100-Pack', 'HDMI Cable 3m', 'DisplayPort Cable 2m',
        'USB Flash Drive 128GB', 'Memory Card Reader', 'Laptop Cooling Pad',
        'Desk Lamp LED', 'Mouse Pad XL', 'Wrist Rest Gel', 'Monitor Arm Dual',
        'Webcam Light Ring', 'Microphone USB Condenser', 'Headphone Stand',
        'Privacy Screen Filter', 'Tablet Stylus Pen', 'Phone Holder Adjustable',
    ]

    for i, product in enumerate(products, 2):
        ws2.cell(row=i, column=1, value=product)

    # Create named range 'ProductList' referring to Products.$A$2:$A$50
    from openpyxl.workbook.defined_name import DefinedName
    defn = DefinedName('ProductList', attr_text="Products!$A$2:$A$50")
    wb.defined_names.add(defn)

    # --- Sheet 3: Instructions (helpful context) ---
    ws3 = wb.create_sheet('Instructions')
    ws3.cell(row=1, column=1, value='Sales Order Entry Instructions')
    ws3.cell(row=1, column=1).font = Font(size=14, bold=True)
    ws3.merge_cells('A1:D1')

    instructions = [
        ('Step 1:', 'Enter the 6-character Customer ID in column B'),
        ('Step 2:', 'Select the order date in column C (must be today or later)'),
        ('Step 3:', 'Choose a product from the dropdown in column D'),
        ('Step 4:', 'Enter the quantity (1-9999) in column E'),
        ('Step 5:', 'Enter the discount percentage (0-50%) in column F'),
        ('Step 6:', 'Select the payment method in column G'),
    ]
    for r, (step, desc) in enumerate(instructions, 3):
        ws3.cell(row=r, column=1, value=step).font = Font(bold=True)
        ws3.cell(row=r, column=2, value=desc)

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 55

    # No data validations on any cells — that is the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
