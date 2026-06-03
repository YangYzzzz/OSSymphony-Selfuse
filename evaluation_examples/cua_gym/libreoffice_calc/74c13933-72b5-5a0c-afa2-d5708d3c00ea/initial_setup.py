"""
Initial Setup: Freeze first column and first row on 'Sales Matrix' sheet
Task ID: calc_ps_056
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Sales Matrix ---
    ws = wb.active
    ws.title = 'Sales Matrix'

    # Headers
    headers = ['Region', 'Q1', 'Q2', 'Q3', 'Q4']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows 2-20 (19 regions)
    regions = [
        'Northeast',
        'Southeast',
        'Midwest',
        'Southwest',
        'Pacific Northwest',
        'Mountain West',
        'Great Plains',
        'Mid-Atlantic',
        'New England',
        'Gulf Coast',
        'Central Valley',
        'Appalachia',
        'Upper Midwest',
        'Deep South',
        'Northern Plains',
        'Inland Empire',
        'Cascadia',
        'Piedmont',
        'Ozarks',
    ]

    sales_data = [
        [142500, 158300, 167200, 189400],
        [98700, 105400, 112800, 121600],
        [176300, 182100, 195600, 204800],
        [87400, 93200, 101500, 108700],
        [134600, 141800, 152300, 163500],
        [65200, 71800, 78400, 85100],
        [112400, 119600, 128300, 137500],
        [203500, 215800, 228400, 241600],
        [156800, 164200, 173500, 182900],
        [91300, 98700, 106200, 114800],
        [78600, 84200, 91500, 98300],
        [54300, 59800, 65200, 71400],
        [103200, 110500, 118700, 126400],
        [72800, 78400, 85600, 92100],
        [48900, 53600, 58200, 63800],
        [89100, 95400, 102700, 110300],
        [121700, 128900, 136400, 145200],
        [67500, 73200, 79800, 86400],
        [43200, 47800, 52600, 57900],
    ]

    for r, (region, data) in enumerate(zip(regions, sales_data), 2):
        ws.cell(row=r, column=1, value=region)
        for c, val in enumerate(data, 2):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 14

    # NO freeze panes - this is the task for the agent to do
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
