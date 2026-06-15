"""
Reward Script: Freeze first column and first row on 'Sales Matrix' sheet
Task ID: calc_ps_056
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): freeze_panes is set (not None) — some freeze applied
  Component 2 (0.6): freeze_panes is exactly 'B2' — correct freeze for row 1 + column A
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_056'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Sales Matrix' sheet must exist
    if 'Sales Matrix' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Sales Matrix' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sales Matrix']

    # Component 1: Freeze panes is set (not None) — some freeze applied (0.4 points)
    try:
        freeze = ws.freeze_panes
        if freeze is not None:
            print(f"PASS: Component 1 — freeze_panes is set to '{freeze}' (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — freeze_panes is None (no freeze applied)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Freeze panes is exactly 'B2' — freezes row 1 AND column A (0.6 points)
    try:
        freeze = ws.freeze_panes
        if freeze is not None and str(freeze) == 'B2':
            print(f"PASS: Component 2 — freeze_panes is exactly 'B2' (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 2 — expected freeze_panes='B2', found '{freeze}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
