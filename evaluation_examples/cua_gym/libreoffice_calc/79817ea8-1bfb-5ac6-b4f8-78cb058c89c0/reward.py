"""
Reward Script: 3-color scale conditional formatting on KPI data
Task ID: calc_gfl_082
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): CF rule exists on B2:E35
  Component 2 (0.20): Rule is a 3-color scale (colorScale with 3 colors)
  Component 3 (0.20): Min color is red (FFFF0000) with cfvo type=min
  Component 4 (0.15): Midpoint color is yellow (FFFFFF00) with cfvo type=percentile
  Component 5 (0.15): Max color is green (FF00FF00) with cfvo type=max
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_082'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify 3-color scale conditional formatting on B2:E35.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the KPIs sheet
    if 'KPIs' not in wb.sheetnames:
        print("FAIL: Sheet 'KPIs' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['KPIs']
    cf_list = list(ws.conditional_formatting)

    # Component 1: CF rule exists covering B2:E35 (0.30 points)
    try:
        # Find any CF rule whose range covers B2:E35
        target_range_found = False
        matching_cf = None
        matching_rule = None

        for cf in cf_list:
            cf_range_str = str(cf).strip()
            # Accept if the range string contains B2:E35 or is equivalent
            # Parse the ranges in the CF object
            for cell_range in cf.cells.ranges:
                r_str = str(cell_range)
                # Check if B2:E35 is covered
                if r_str == 'B2:E35':
                    target_range_found = True
                    matching_cf = cf
                    break
            if target_range_found:
                break

        # Also check if it covers B2:E35 even if expressed differently
        if not target_range_found:
            for cf in cf_list:
                for cell_range in cf.cells.ranges:
                    # Check if range starts at B2 and ends at E35
                    if (cell_range.min_row == 2 and cell_range.max_row == 35 and
                        cell_range.min_col == 2 and cell_range.max_col == 5):
                        target_range_found = True
                        matching_cf = cf
                        break
                if target_range_found:
                    break

        if target_range_found:
            print(f"PASS: Component 1 - CF rule found covering B2:E35 (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 - No CF rule covering B2:E35 found. CF count: {len(cf_list)}")
            for cf in cf_list:
                print(f"  Found CF range: {cf}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Find the colorScale rule for further checks
    color_scale_rule = None
    if matching_cf:
        for rule in matching_cf.rules:
            if rule.type == 'colorScale' and hasattr(rule, 'colorScale') and rule.colorScale:
                color_scale_rule = rule
                break
    else:
        # Search all CF rules for any colorScale
        for cf in cf_list:
            for rule in cf.rules:
                if rule.type == 'colorScale' and hasattr(rule, 'colorScale') and rule.colorScale:
                    color_scale_rule = rule
                    break
            if color_scale_rule:
                break

    # Component 2: Rule is a 3-color scale (0.20 points)
    try:
        if color_scale_rule and color_scale_rule.colorScale:
            cs = color_scale_rule.colorScale
            num_colors = len(cs.color)
            num_cfvo = len(cs.cfvo)
            if num_colors == 3 and num_cfvo == 3:
                print(f"PASS: Component 2 - 3-color scale rule found ({num_colors} colors, {num_cfvo} cfvo) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 - Expected 3-color scale, got {num_colors} colors and {num_cfvo} cfvo")
        else:
            print(f"FAIL: Component 2 - No colorScale rule found")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Min color is red with cfvo type=min (0.20 points)
    try:
        if color_scale_rule and color_scale_rule.colorScale:
            cs = color_scale_rule.colorScale
            if len(cs.color) >= 3 and len(cs.cfvo) >= 3:
                min_color = cs.color[0].rgb
                min_type = cs.cfvo[0].type
                # Accept various red shades
                is_red = min_color in ('FFFF0000', 'FFF8696B', 'FFFF4444', 'FFCC0000')
                # More lenient: check if R channel is high, G and B are low
                if not is_red and min_color and len(min_color) == 8:
                    r = int(min_color[2:4], 16)
                    g = int(min_color[4:6], 16)
                    b = int(min_color[6:8], 16)
                    is_red = (r >= 180 and g < 130 and b < 130)

                if is_red and min_type == 'min':
                    print(f"PASS: Component 3 - Min color is red ({min_color}), cfvo type=min (0.20 pts)")
                    total_score += 0.20
                else:
                    print(f"FAIL: Component 3 - Min: color={min_color} (red={is_red}), type={min_type}")
            else:
                print(f"FAIL: Component 3 - Not enough colors/cfvo in colorScale")
        else:
            print(f"FAIL: Component 3 - No colorScale rule")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Midpoint color is yellow with cfvo type=percentile (0.15 points)
    try:
        if color_scale_rule and color_scale_rule.colorScale:
            cs = color_scale_rule.colorScale
            if len(cs.color) >= 3 and len(cs.cfvo) >= 3:
                mid_color = cs.color[1].rgb
                mid_type = cs.cfvo[1].type
                # Accept yellow shades
                is_yellow = mid_color in ('FFFFFF00', 'FFFFEB84', 'FFFFD700', 'FFFFCC00')
                if not is_yellow and mid_color and len(mid_color) == 8:
                    r = int(mid_color[2:4], 16)
                    g = int(mid_color[4:6], 16)
                    b = int(mid_color[6:8], 16)
                    is_yellow = (r >= 200 and g >= 180 and b < 130)

                # Accept percentile or percent for midpoint type
                is_mid_type = mid_type in ('percentile', 'percent', 'num')

                if is_yellow and is_mid_type:
                    print(f"PASS: Component 4 - Midpoint color is yellow ({mid_color}), cfvo type={mid_type} (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 4 - Mid: color={mid_color} (yellow={is_yellow}), type={mid_type}")
            else:
                print(f"FAIL: Component 4 - Not enough colors/cfvo in colorScale")
        else:
            print(f"FAIL: Component 4 - No colorScale rule")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Max color is green with cfvo type=max (0.15 points)
    try:
        if color_scale_rule and color_scale_rule.colorScale:
            cs = color_scale_rule.colorScale
            if len(cs.color) >= 3 and len(cs.cfvo) >= 3:
                max_color = cs.color[2].rgb
                max_type = cs.cfvo[2].type
                # Accept green shades
                is_green = max_color in ('FF00FF00', 'FF63BE7B', 'FF00B050', 'FF008000', 'FF00CC00')
                if not is_green and max_color and len(max_color) == 8:
                    r = int(max_color[2:4], 16)
                    g = int(max_color[4:6], 16)
                    b = int(max_color[6:8], 16)
                    is_green = (g >= 150 and r < 130 and b < 150)

                if is_green and max_type == 'max':
                    print(f"PASS: Component 5 - Max color is green ({max_color}), cfvo type=max (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 5 - Max: color={max_color} (green={is_green}), type={max_type}")
            else:
                print(f"FAIL: Component 5 - Not enough colors/cfvo in colorScale")
        else:
            print(f"FAIL: Component 5 - No colorScale rule")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
