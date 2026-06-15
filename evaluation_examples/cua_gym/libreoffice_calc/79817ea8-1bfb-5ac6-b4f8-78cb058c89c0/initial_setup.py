"""
Initial Setup: KPI matrix spreadsheet with quarterly scores
Task ID: calc_gfl_082
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KPIs'

    # Headers
    headers = ['KPI Name', 'Q1 Score', 'Q2 Score', 'Q3 Score', 'Q4 Score']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 34 realistic KPI metrics with scores between 0 and 100
    kpi_data = [
        ['Customer Satisfaction Index', 78, 82, 85, 88],
        ['Net Promoter Score', 45, 52, 48, 55],
        ['Employee Retention Rate', 92, 89, 91, 93],
        ['Revenue Growth Rate', 65, 71, 68, 74],
        ['Operating Margin', 34, 38, 36, 41],
        ['Customer Acquisition Cost', 58, 53, 49, 45],
        ['Monthly Active Users', 72, 76, 81, 85],
        ['Average Response Time', 88, 85, 90, 92],
        ['First Call Resolution', 67, 70, 73, 75],
        ['Inventory Turnover', 55, 58, 62, 60],
        ['Order Fulfillment Rate', 94, 96, 95, 97],
        ['Website Conversion Rate', 23, 27, 31, 35],
        ['Social Media Engagement', 41, 48, 52, 56],
        ['Product Defect Rate', 12, 9, 7, 5],
        ['Training Completion Rate', 76, 82, 88, 91],
        ['Budget Variance', 85, 80, 78, 83],
        ['On-Time Delivery Rate', 89, 91, 93, 95],
        ['Employee Productivity Index', 71, 74, 77, 80],
        ['Customer Lifetime Value', 63, 68, 72, 76],
        ['Market Share Growth', 29, 33, 37, 42],
        ['IT System Uptime', 99, 98, 99, 100],
        ['Quality Assurance Score', 82, 85, 87, 90],
        ['Supplier Performance Index', 73, 76, 79, 81],
        ['Cash Flow Efficiency', 61, 65, 63, 67],
        ['R&D Pipeline Value', 44, 50, 55, 59],
        ['Sales Quota Attainment', 87, 83, 90, 92],
        ['Cross-Sell Revenue Ratio', 36, 40, 43, 47],
        ['Customer Churn Rate', 15, 12, 10, 8],
        ['Process Cycle Time', 68, 72, 75, 79],
        ['Environmental Compliance', 95, 96, 97, 98],
        ['Data Accuracy Rate', 88, 91, 93, 95],
        ['Innovation Index', 51, 56, 61, 65],
        ['Workforce Diversity Score', 64, 68, 71, 74],
        ['Risk Management Effectiveness', 77, 80, 83, 86],
    ]

    for r, row_data in enumerate(kpi_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    # No conditional formatting - that's the task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
