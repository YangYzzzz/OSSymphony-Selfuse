"""
Reward Script: Merge cells A1:F1 and format with centered text, bold 16pt Arial font,
               gradient-style solid blue (#4472C4) background, and white font color.
Task ID: calc_fmt_cell_merge_format_091
Domain: libreoffice_calc
Scoring:
  Component 1: Cells A1:F1 are merged (0.30 pts)
  Component 2: Font is bold, size=16pt, name=Arial (0.30 pts)
  Component 3: Horizontal alignment is center (0.15 pts)
  Component 4: Background fill is solid #4472C4 (0.15 pts)
  Component 5: Font color is white #FFFFFF (0.10 pts)
  Total: 1.00
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_cell_merge_format_091'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Checks that A1:F1 is merged and the merged cell has the required formatting.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'Department Report' must exist
    if 'Department Report' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Department Report' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Department Report']
    cell_a1 = ws['A1']

    # Component 1: Cells A1:F1 are merged (0.30 points)
    # Checks that the merged range A1:F1 exists in the worksheet.
    # On the initial file, there are no merged ranges, so this FAILS on initial.
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        if 'A1:F1' in merged_ranges:
            print(f"PASS: Component 1 — Cells A1:F1 are merged (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected merged range A1:F1, found: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Font is bold, size=16pt, font name=Arial (0.30 points)
    # On the initial file, cell A1 has bold=False, size=11, name=Calibri.
    # This FAILS on initial and PASSES on golden.
    try:
        font = cell_a1.font
        is_bold = font.bold == True
        is_size_16 = font.size == 16.0
        is_arial = font.name == 'Arial'

        if is_bold and is_size_16 and is_arial:
            print(f"PASS: Component 2 — Font is bold={font.bold}, size={font.size}, name={font.name} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — Expected bold=True, size=16, name=Arial; "
                  f"found bold={font.bold}, size={font.size}, name={font.name}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Horizontal alignment is center (0.15 points)
    # On the initial file, alignment is None. This FAILS on initial.
    try:
        align = cell_a1.alignment.horizontal
        if align == 'center':
            print(f"PASS: Component 3 — Horizontal alignment is '{align}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Expected horizontal='center', found: {align}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Background fill is solid #4472C4 (ARGB: FF4472C4) (0.15 points)
    # On the initial file, fill is None/empty (00000000). This FAILS on initial.
    try:
        fg_color = cell_a1.fill.fgColor.rgb
        fill_type = cell_a1.fill.patternType
        # Accept both FF4472C4 (correct ARGB) and 004472C4 (may occur if set with 6-char)
        expected_color = 'FF4472C4'
        if fill_type == 'solid' and fg_color.upper() == expected_color.upper():
            print(f"PASS: Component 4 — Background fill is solid {fg_color} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Expected solid fill FF4472C4, "
                  f"found fill_type={fill_type}, fgColor={fg_color}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Font color is white #FFFFFF (ARGB: FFFFFFFF) (0.10 points)
    # On the initial file, font color is unset (default). This FAILS on initial.
    try:
        font_color_rgb = cell_a1.font.color.rgb
        # Accept both FFFFFFFF (full ARGB with alpha) and 00FFFFFF
        if font_color_rgb.upper() in ('FFFFFFFF', '00FFFFFF'):
            print(f"PASS: Component 5 — Font color is {font_color_rgb} (white) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Expected font color FFFFFFFF (white), found: {font_color_rgb}")
    except Exception as e:
        print(f"ERROR: Component 5 — font color error: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
