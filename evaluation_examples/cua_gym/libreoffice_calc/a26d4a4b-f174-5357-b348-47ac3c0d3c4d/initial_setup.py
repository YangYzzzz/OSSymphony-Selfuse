"""
Initial Setup: Merge cells A1:F1 and format the merged cell
Task ID: calc_fmt_cell_merge_format_091
Domain: libreoffice_calc

Creates a spreadsheet with a Department Report sheet.
Row 1 contains a title in A1 (not merged, no special formatting).
Rows 2+ contain realistic HR employee data.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_cell_merge_format_091'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Department Report ---
    ws = wb.active
    ws.title = 'Department Report'

    # Row 1: Title in A1 only (NOT merged, NO special formatting)
    ws['A1'] = 'Human Resources Department \u2014 Q1 2025 Report'
    # B1:F1 are intentionally empty — NOT merged

    # Row 2: Column headers
    headers = ['Employee ID', 'Full Name', 'Job Title', 'Department', 'Salary (USD)', 'Start Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
        ws.cell(row=2, column=col).font = Font(bold=True)

    # Rows 3-17: Realistic HR employee data
    data = [
        ['EMP-1001', 'Sarah Chen',        'HR Manager',          'Human Resources', 92000,  '2021-03-15'],
        ['EMP-1002', 'Marcus Johnson',     'Recruiter',           'Human Resources', 64500,  '2022-07-01'],
        ['EMP-1003', 'Diana Okonkwo',      'Payroll Specialist',  'Human Resources', 71000,  '2020-11-20'],
        ['EMP-1004', 'Liam Fitzgerald',    'Benefits Coordinator','Human Resources', 58000,  '2023-01-10'],
        ['EMP-1005', 'Priya Sharma',       'Training Manager',    'Human Resources', 85000,  '2019-05-22'],
        ['EMP-1006', 'Carlos Rivera',      'HR Analyst',          'Human Resources', 68000,  '2021-09-14'],
        ['EMP-1007', 'Aisha Patel',        'Talent Acquisition',  'Human Resources', 74000,  '2022-03-28'],
        ['EMP-1008', 'Thomas Nguyen',      'HR Generalist',       'Human Resources', 62000,  '2023-06-05'],
        ['EMP-1009', 'Sophie Moreau',      'Compensation Analyst','Human Resources', 79000,  '2020-08-17'],
        ['EMP-1010', 'Daniel Park',        'HRIS Specialist',     'Human Resources', 76500,  '2021-12-01'],
        ['EMP-1011', 'Fatima Al-Hassan',   'Employee Relations',  'Human Resources', 69000,  '2022-10-11'],
        ['EMP-1012', 'Noah Williams',      'HR Coordinator',      'Human Resources', 55000,  '2024-01-08'],
        ['EMP-1013', 'Isabella Torres',    'Learning Specialist',  'Human Resources', 72000,  '2020-04-30'],
        ['EMP-1014', 'Ethan Blackwood',    'Recruiter',           'Human Resources', 65000,  '2023-08-19'],
        ['EMP-1015', 'Mei-Ling Zhang',     'HR Director',         'Human Resources', 115000, '2018-02-14'],
    ]
    for r, row_data in enumerate(data, 3):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    # Set header row height
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
