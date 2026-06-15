"""
Initial Setup: HR Analytics Dashboard - Create data sheets for charting
Task ID: calc_hr_095
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 'Data' ---
    ws = wb.active
    ws.title = 'Data'

    # Section 1: Monthly Hiring & Turnover (A:C, rows 1-13)
    headers_ac = ['Month', 'Hires', 'Terminations']
    for col, h in enumerate(headers_ac, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    monthly_data = [
        ['Jan 2025', 12, 5],
        ['Feb 2025', 8, 3],
        ['Mar 2025', 15, 7],
        ['Apr 2025', 10, 4],
        ['May 2025', 18, 6],
        ['Jun 2025', 14, 8],
        ['Jul 2025', 9, 5],
        ['Aug 2025', 11, 3],
        ['Sep 2025', 16, 9],
        ['Oct 2025', 13, 7],
        ['Nov 2025', 7, 4],
        ['Dec 2025', 20, 10],
    ]
    for r, row_data in enumerate(monthly_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Section 2: Headcount by Department (E:F, rows 1-6)
    dept_headers = ['Department', 'Headcount']
    for col_offset, h in enumerate(dept_headers):
        cell = ws.cell(row=1, column=5 + col_offset, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    dept_data = [
        ['Engineering', 145],
        ['Sales', 89],
        ['Marketing', 52],
        ['Human Resources', 34],
        ['Finance', 41],
    ]
    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data):
            ws.cell(row=r, column=5 + c, value=val)

    # Section 3: Salary Distribution (H:I, rows 1-7)
    salary_headers = ['Salary Range', 'Employee Count']
    for col_offset, h in enumerate(salary_headers):
        cell = ws.cell(row=1, column=8 + col_offset, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    salary_data = [
        ['$30k-$50k', 42],
        ['$50k-$70k', 78],
        ['$70k-$90k', 115],
        ['$90k-$110k', 67],
        ['$110k-$130k', 38],
        ['$130k+', 21],
    ]
    for r, row_data in enumerate(salary_data, 2):
        for c, val in enumerate(row_data):
            ws.cell(row=r, column=8 + c, value=val)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16

    # --- Sheet 'Dashboard' (empty, ready for charts) ---
    wb.create_sheet('Dashboard')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
