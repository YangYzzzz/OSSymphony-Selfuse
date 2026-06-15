"""
Initial Setup: Risk Register spreadsheet with risk scores (no conditional formatting)
Task ID: calc_fmt_condfmt_colorscale_3color_046
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_condfmt_colorscale_3color_046'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Risk Register sheet ---
    ws = wb.active
    ws.title = 'Risk Register'

    # Headers
    headers = ['Risk ID', 'Description', 'Risk Score', 'Owner']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Realistic risk data — 24 rows (rows 2-25)
    risk_data = [
        ('RK-001', 'Data breach via unpatched server vulnerabilities', 9, 'Alice Chen'),
        ('RK-002', 'Critical vendor contract expiry with no backup supplier', 8, 'Marcus Johnson'),
        ('RK-003', 'Regulatory non-compliance with GDPR data handling', 7, 'Priya Sharma'),
        ('RK-004', 'Cloud infrastructure outage during peak trading hours', 9, 'David Lee'),
        ('RK-005', 'Key personnel turnover in engineering department', 6, 'Sarah Mitchell'),
        ('RK-006', 'Phishing attacks targeting finance team', 8, 'Tom Ramirez'),
        ('RK-007', 'Delayed product launch due to supply chain disruption', 5, 'Lisa Wong'),
        ('RK-008', 'Failure of backup power systems at data center', 7, 'James Park'),
        ('RK-009', 'Software license compliance violations', 4, 'Elena Vasquez'),
        ('RK-010', 'Inadequate disaster recovery plan for core systems', 8, 'Michael Brown'),
        ('RK-011', 'Project cost overrun exceeding 20% of budget', 5, 'Natasha Ivanova'),
        ('RK-012', 'Loss of customer data due to database corruption', 9, 'Chris Taylor'),
        ('RK-013', 'Fraudulent access to financial reporting systems', 7, 'Amanda Foster'),
        ('RK-014', 'Underperformance of third-party logistics partner', 3, 'Kevin Nguyen'),
        ('RK-015', 'Insider threat from privileged system administrator', 6, 'Rachel Kim'),
        ('RK-016', 'Non-renewal of key software maintenance agreements', 4, 'Omar Hassan'),
        ('RK-017', 'Cybersecurity incident affecting payment gateway', 10, 'Sophia Martinez'),
        ('RK-018', 'Inadequate staff training on new ERP system', 3, 'Daniel Cooper'),
        ('RK-019', 'Merger integration delays creating operational gaps', 6, 'Yuki Tanaka'),
        ('RK-020', 'Environmental regulation changes impacting operations', 5, 'Brendan Walsh'),
        ('RK-021', 'Hardware failure in production manufacturing line', 7, 'Ingrid Olsen'),
        ('RK-022', 'Intellectual property theft by competitor', 8, 'Pablo Garcia'),
        ('RK-023', 'Prolonged network outage in regional office', 2, 'Fatima Al-Amin'),
        ('RK-024', 'Reputational damage from social media incident', 1, 'Jason White'),
    ]

    for r, row_data in enumerate(risk_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20

    # No conditional formatting — task requires the agent to add it

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
