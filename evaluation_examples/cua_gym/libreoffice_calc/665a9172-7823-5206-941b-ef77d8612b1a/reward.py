"""
Reward Script: Set up a 3-color scale on risk scores in C2:C25
Task ID: calc_fmt_condfmt_colorscale_3color_046
Domain: libreoffice_calc
Scoring:
  Component 1: CF rule exists on range C2:C25 (0.3 pts)
  Component 2: Rule type is colorScale with correct cfvo stops (min/percentile50/max) (0.3 pts)
  Component 3: Colors are correct red/yellow/green in expected positions (0.4 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_condfmt_colorscale_3color_046'

# Expected colors (ARGB format as stored by openpyxl)
EXPECTED_MIN_COLOR = 'FFFF0000'    # red for lowest values
EXPECTED_MID_COLOR = 'FFFFFF00'    # yellow for mid-range (50th percentile)
EXPECTED_MAX_COLOR = 'FF00FF00'    # green for highest values

# Expected CF range
EXPECTED_CF_RANGE = 'C2:C25'


def normalize_color(rgb_str):
    """Normalize color string to 8-char ARGB, handling both 6-char and 8-char inputs."""
    if rgb_str is None:
        return None
    rgb_str = str(rgb_str).upper().strip()
    if len(rgb_str) == 6:
        return 'FF' + rgb_str
    return rgb_str


def colors_match(actual_rgb, expected_argb):
    """Check if actual color matches expected ARGB, handling format variations."""
    if actual_rgb is None:
        return False
    actual_norm = normalize_color(actual_rgb)
    expected_norm = normalize_color(expected_argb)
    return actual_norm == expected_norm


def verify_task(file_path):
    """
    Verify that a 3-color scale conditional formatting rule was applied to C2:C25.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the sheet
    try:
        if 'Risk Register' not in wb.sheetnames:
            print(f"CRITICAL: Sheet 'Risk Register' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0
        ws = wb['Risk Register']
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet 'Risk Register': {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A CF rule exists on range C2:C25 (0.3 points)
    # This FAILS on initial (no CF) and PASSES on golden (CF exists on C2:C25)
    try:
        cf_list = list(ws.conditional_formatting)
        target_cf = None

        for cf_range in cf_list:
            # Check if this CF covers C2:C25
            cf_str = str(cf_range)
            if 'C2:C25' in cf_str:
                target_cf = cf_range
                break

        if target_cf is not None:
            print(f"PASS: Component 1 — CF rule found on range C2:C25 (0.3 pts)")
            total_score += 0.3
        else:
            cf_ranges_str = [str(r) for r in cf_list]
            print(f"FAIL: Component 1 — No CF rule found on C2:C25. Found ranges: {cf_ranges_str}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Rule type is colorScale with correct cfvo stops (0.3 points)
    # cfvo[0]: type=min, cfvo[1]: type=percentile at 50, cfvo[2]: type=max
    # This FAILS on initial (no CF at all) and PASSES on golden
    try:
        if target_cf is None:
            print("FAIL: Component 2 — No CF rule to inspect (prerequisite Component 1 failed)")
        else:
            rules_list = list(ws.conditional_formatting[target_cf])
            colorscale_rule = None
            for rule in rules_list:
                if rule.type == 'colorScale' and rule.colorScale is not None:
                    colorscale_rule = rule
                    break

            if colorscale_rule is None:
                rule_types = [r.type for r in rules_list]
                print(f"FAIL: Component 2 — No colorScale rule found. Rule types: {rule_types}")
            else:
                cs = colorscale_rule.colorScale
                cfvo_list = cs.cfvo
                if len(cfvo_list) == 3:
                    cfvo0_ok = (cfvo_list[0].type == 'min')
                    cfvo1_ok = (cfvo_list[1].type == 'percentile' and
                                cfvo_list[1].val is not None and
                                abs(float(cfvo_list[1].val) - 50.0) < 0.1)
                    cfvo2_ok = (cfvo_list[2].type == 'max')

                    if cfvo0_ok and cfvo1_ok and cfvo2_ok:
                        print(f"PASS: Component 2 — colorScale rule with correct cfvo stops: "
                              f"min/percentile(50)/max (0.3 pts)")
                        total_score += 0.3
                    else:
                        print(f"FAIL: Component 2 — colorScale cfvo stops incorrect. "
                              f"cfvo[0].type={cfvo_list[0].type} (want min), "
                              f"cfvo[1].type={cfvo_list[1].type} val={cfvo_list[1].val} (want percentile/50), "
                              f"cfvo[2].type={cfvo_list[2].type} (want max)")
                else:
                    print(f"FAIL: Component 2 — Expected 3 cfvo stops, found {len(cfvo_list)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Colors are red (min) / yellow (mid) / green (max) (0.4 points)
    # Exact colors: FFFF0000 (red), FFFFFF00 (yellow), FF00FF00 (green)
    # This FAILS on initial (no CF) and PASSES on golden (correct colors set)
    try:
        if target_cf is None:
            print("FAIL: Component 3 — No CF rule to inspect (prerequisite Component 1 failed)")
        else:
            rules_list = list(ws.conditional_formatting[target_cf])
            colorscale_rule = None
            for rule in rules_list:
                if rule.type == 'colorScale' and rule.colorScale is not None:
                    colorscale_rule = rule
                    break

            if colorscale_rule is None:
                print("FAIL: Component 3 — No colorScale rule found")
            else:
                cs = colorscale_rule.colorScale
                color_list = cs.color
                if len(color_list) == 3:
                    c0_rgb = color_list[0].rgb
                    c1_rgb = color_list[1].rgb
                    c2_rgb = color_list[2].rgb

                    min_ok = colors_match(c0_rgb, EXPECTED_MIN_COLOR)
                    mid_ok = colors_match(c1_rgb, EXPECTED_MID_COLOR)
                    max_ok = colors_match(c2_rgb, EXPECTED_MAX_COLOR)

                    if min_ok and mid_ok and max_ok:
                        print(f"PASS: Component 3 — Colors correct: "
                              f"min=red({c0_rgb}), mid=yellow({c1_rgb}), max=green({c2_rgb}) (0.4 pts)")
                        total_score += 0.4
                    else:
                        # Award partial credit if some colors are correct
                        partial = 0.0
                        if min_ok:
                            partial += 0.15
                            print(f"  PASS: min color is red ({c0_rgb})")
                        else:
                            print(f"  FAIL: min color expected {EXPECTED_MIN_COLOR}, got {c0_rgb}")
                        if mid_ok:
                            partial += 0.10
                            print(f"  PASS: mid color is yellow ({c1_rgb})")
                        else:
                            print(f"  FAIL: mid color expected {EXPECTED_MID_COLOR}, got {c1_rgb}")
                        if max_ok:
                            partial += 0.15
                            print(f"  PASS: max color is green ({c2_rgb})")
                        else:
                            print(f"  FAIL: max color expected {EXPECTED_MAX_COLOR}, got {c2_rgb}")
                        if partial > 0:
                            print(f"PARTIAL: Component 3 — Partial color match ({partial} pts)")
                            total_score += partial
                        else:
                            print(f"FAIL: Component 3 — All colors incorrect")
                else:
                    print(f"FAIL: Component 3 — Expected 3 colors, found {len(color_list)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
