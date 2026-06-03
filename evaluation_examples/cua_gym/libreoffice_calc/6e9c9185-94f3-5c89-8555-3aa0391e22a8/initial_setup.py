"""
Initial Setup: Copenhagen restaurant directory lookup task
Task ID: osworld_multi_apps_restaurant_lookup_015
Domain: libreoffice_calc (multi-app: gedit + LibreOffice Calc)

Creates:
  - /home/user/Desktop/copenhagen_restaurants.txt  (list of 5 restaurant names, unsorted)
  - /home/user/CPH_DIRECTORY.xlsx  (spreadsheet with Name column pre-filled, unsorted, no contact data)

Opens:
  - gedit with copenhagen_restaurants.txt
  - LibreOffice Calc with CPH_DIRECTORY.xlsx
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_015'
TXT_PATH = f'{DESKTOP}/copenhagen_restaurants.txt'
XLSX_PATH = f'{WORKDIR}/CPH_DIRECTORY.xlsx'


def launch_gui(command: str, delay_sec: float = 1.5):
    """Launch a GUI app on the VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_txt():
    """Create the restaurant names text file on the Desktop (unsorted order)."""
    os.makedirs(DESKTOP, exist_ok=True)
    content = (
        "Noma\n"
        "Geranium\n"
        "Kadeau\n"
        "Amass\n"
        "AOC\n"
    )
    with open(TXT_PATH, 'w') as f:
        f.write(content)
    print(f'Text file created: {TXT_PATH}')


def create_xlsx():
    """Create CPH_DIRECTORY.xlsx with Name column pre-filled (unsorted), other columns empty."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Header row
    headers = ['Name', 'Address', 'Website', 'Phone']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20

    # Restaurant names (unsorted — matches the txt file order)
    # Rows A2:A6, Address/Website/Phone columns left empty (task requires lookup)
    names_unsorted = [
        'Noma',
        'Geranium',
        'Kadeau',
        'Amass',
        'AOC',
    ]
    for row_idx, name in enumerate(names_unsorted, 2):
        ws.cell(row=row_idx, column=1, value=name)
        # Columns B, C, D intentionally empty — agent must fill these

    wb.save(XLSX_PATH)
    print(f'Spreadsheet created: {XLSX_PATH}')


def main():
    create_txt()
    create_xlsx()

    # GUI-ready startup: open gedit with the txt file, then LibreOffice Calc
    launch_gui(f'gedit "{TXT_PATH}"', delay_sec=2.0)
    launch_gui(f'libreoffice --calc "{XLSX_PATH}"', delay_sec=2.5)

    print('GUI_READY: launched gedit and LibreOffice Calc with DISPLAY=:0')


main()
