"""
Reward Script: Copenhagen Restaurant Directory Lookup
Task ID: osworld_multi_apps_restaurant_lookup_015
Domain: libreoffice_calc

Task: Search Google Maps for each Copenhagen restaurant's address, website,
      and phone number, fill the data into CPH_DIRECTORY.xlsx, and sort
      the rows alphabetically by restaurant name.

Scoring Rubric:
  Component 1: Rows sorted alphabetically by restaurant name  — 0.30 pts
  Component 2: Address fields (B2:B6) all filled              — 0.30 pts
  Component 3: Website fields (C2:C6) all filled              — 0.20 pts
  Component 4: Phone fields (D2:D6) all filled                — 0.20 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_restaurant_lookup_015'

# Expected sorted order (A-Z) for 5 Copenhagen restaurants
EXPECTED_SORTED_NAMES = ['AOC', 'Amass', 'Geranium', 'Kadeau', 'Noma']
EXPECTED_RESTAURANT_SET = set(EXPECTED_SORTED_NAMES)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Checks (in order):
      1. Rows sorted alphabetically by name (AOC, Amass, Geranium, Kadeau, Noma)
      2. Address column (B) filled for all 5 restaurants
      3. Website column (C) filled for all 5 restaurants
      4. Phone column (D) filled for all 5 restaurants
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Verify we have exactly 5 data rows with the expected restaurant names
    try:
        data_rows = []
        for row_idx in range(2, 7):  # rows 2-6
            name_val = ws.cell(row=row_idx, column=1).value
            addr_val = ws.cell(row=row_idx, column=2).value
            web_val  = ws.cell(row=row_idx, column=3).value
            phone_val= ws.cell(row=row_idx, column=4).value
            data_rows.append({
                'name': name_val,
                'address': addr_val,
                'website': web_val,
                'phone': phone_val,
                'row': row_idx
            })

        actual_names = [r['name'] for r in data_rows]
        actual_name_set = set(n for n in actual_names if n is not None)

        if not EXPECTED_RESTAURANT_SET.issubset(actual_name_set):
            missing = EXPECTED_RESTAURANT_SET - actual_name_set
            print(f"PRECONDITION FAIL: Missing restaurant names: {missing}")
            print("REWARD: 0.0")
            return 0.0

        print(f"PRECONDITION PASS: Found all 5 restaurants in rows 2-6")

    except Exception as e:
        print(f"CRITICAL: Could not read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Rows sorted alphabetically by restaurant name (0.30 points)
    # This checks the task-introduced change: sorting the data A-Z by name.
    # Initial state has names in: Noma, Geranium, Kadeau, Amass, AOC (unsorted).
    # Golden state should have: AOC, Amass, Geranium, Kadeau, Noma (sorted A-Z).
    try:
        actual_names_in_order = [r['name'] for r in data_rows]
        if actual_names_in_order == EXPECTED_SORTED_NAMES:
            print(f"PASS: Component 1 — Rows sorted A-Z: {actual_names_in_order} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — Expected order {EXPECTED_SORTED_NAMES}, found {actual_names_in_order}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Address fields filled for all 5 restaurants (0.30 points)
    # Initial state has all address fields empty (None).
    # Task requires: search and fill address for each restaurant.
    try:
        addresses = [r['address'] for r in data_rows]
        filled_addresses = [a for a in addresses if a is not None and str(a).strip() != '']
        if len(filled_addresses) == 5:
            print(f"PASS: Component 2 — All 5 address fields filled (0.30 pts)")
            total_score += 0.30
        else:
            missing_count = 5 - len(filled_addresses)
            print(f"FAIL: Component 2 — {missing_count} address field(s) missing. "
                  f"Addresses found: {filled_addresses}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Website fields filled for all 5 restaurants (0.20 points)
    # Initial state has all website fields empty (None).
    # Task requires: search and fill website for each restaurant.
    try:
        websites = [r['website'] for r in data_rows]
        filled_websites = [w for w in websites if w is not None and str(w).strip() != '']
        if len(filled_websites) == 5:
            print(f"PASS: Component 3 — All 5 website fields filled (0.20 pts)")
            total_score += 0.20
        else:
            missing_count = 5 - len(filled_websites)
            print(f"FAIL: Component 3 — {missing_count} website field(s) missing. "
                  f"Websites found: {filled_websites}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Phone fields filled for all 5 restaurants (0.20 points)
    # Initial state has all phone fields empty (None).
    # Task requires: search and fill phone number for each restaurant.
    try:
        phones = [r['phone'] for r in data_rows]
        filled_phones = [p for p in phones if p is not None and str(p).strip() != '']
        if len(filled_phones) == 5:
            print(f"PASS: Component 4 — All 5 phone fields filled (0.20 pts)")
            total_score += 0.20
        else:
            missing_count = 5 - len(filled_phones)
            print(f"FAIL: Component 4 — {missing_count} phone field(s) missing. "
                  f"Phones found: {filled_phones}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/CPH_DIRECTORY.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
