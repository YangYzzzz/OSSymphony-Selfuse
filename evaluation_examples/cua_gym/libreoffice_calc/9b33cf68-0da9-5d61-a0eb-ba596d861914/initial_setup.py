"""
Initial Setup: Create spreadsheet with metrics and quarter lookup data for dropdown validation task.
Task ID: calc_gcv_063
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Quarterly_Report ---
    ws1 = wb.active
    ws1.title = 'Quarterly_Report'

    # Headers
    ws1.cell(row=1, column=1, value='Metric')
    ws1.cell(row=1, column=2, value='Quarter')
    ws1.cell(row=1, column=3, value='Month')

    # 29 realistic business metrics in A2:A30
    metrics = [
        'Total Revenue',
        'Net Profit',
        'Gross Margin',
        'Operating Expenses',
        'Customer Acquisition Cost',
        'Monthly Recurring Revenue',
        'Churn Rate',
        'Average Order Value',
        'Customer Lifetime Value',
        'Return on Investment',
        'Employee Headcount',
        'Revenue per Employee',
        'Marketing Spend',
        'Sales Pipeline Value',
        'Conversion Rate',
        'Website Traffic',
        'Support Tickets Resolved',
        'Product Defect Rate',
        'Inventory Turnover',
        'Accounts Receivable',
        'Cash Flow from Operations',
        'Debt-to-Equity Ratio',
        'Working Capital',
        'R&D Expenditure',
        'Customer Satisfaction Score',
        'Net Promoter Score',
        'Market Share Percentage',
        'Brand Awareness Index',
        'Supply Chain Efficiency',
    ]

    for i, metric in enumerate(metrics, 2):
        ws1.cell(row=i, column=1, value=metric)

    # B2:B30 and C2:C30 intentionally left empty (agent must add validations)

    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')
    for col in range(1, 4):
        cell = ws1.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws1.column_dimensions['A'].width = 30
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 14

    # --- Sheet2: Lookup ---
    ws2 = wb.create_sheet('Lookup')

    # Quarter-to-month mapping
    # Q1 in A1:A3
    ws2.cell(row=1, column=1, value='January')
    ws2.cell(row=2, column=1, value='February')
    ws2.cell(row=3, column=1, value='March')

    # Q2 in B1:B3
    ws2.cell(row=1, column=2, value='April')
    ws2.cell(row=2, column=2, value='May')
    ws2.cell(row=3, column=2, value='June')

    # Q3 in C1:C3
    ws2.cell(row=1, column=3, value='July')
    ws2.cell(row=2, column=3, value='August')
    ws2.cell(row=3, column=3, value='September')

    # Q4 in D1:D3
    ws2.cell(row=1, column=4, value='October')
    ws2.cell(row=2, column=4, value='November')
    ws2.cell(row=3, column=4, value='December')

    # Define named ranges for each quarter
    # These named ranges are what INDIRECT() will reference
    dn_q1 = DefinedName('Q1', attr_text="Lookup!$A$1:$A$3")
    dn_q2 = DefinedName('Q2', attr_text="Lookup!$B$1:$B$3")
    dn_q3 = DefinedName('Q3', attr_text="Lookup!$C$1:$C$3")
    dn_q4 = DefinedName('Q4', attr_text="Lookup!$D$1:$D$3")

    wb.defined_names.add(dn_q1)
    wb.defined_names.add(dn_q2)
    wb.defined_names.add(dn_q3)
    wb.defined_names.add(dn_q4)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
