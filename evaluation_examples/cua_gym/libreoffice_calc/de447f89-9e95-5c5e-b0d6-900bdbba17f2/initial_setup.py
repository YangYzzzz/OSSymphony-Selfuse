"""
Initial Setup: Highlight duplicate Order IDs with conditional formatting
Task ID: calc_dop_dedup_highlight_068
Domain: libreoffice_calc

Creates an OrderLog sheet with 119 order records in rows 2-120.
Column A has 14 Order IDs that appear more than once.
No conditional formatting applied (that is the task).
"""

import openpyxl
from openpyxl.styles import Font
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_dedup_highlight_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OrderLog'

    # --- Headers ---
    headers = ['Order ID', 'Date', 'Customer', 'Amount']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # Build order records: 119 rows, 14 duplicate Order IDs
    # The 14 duplicated IDs and their row positions (as per context):
    # ORD-2234 appears in rows 5, 45, 89 (3 occurrences = 2 duplicates extra)
    # ORD-2891 appears in rows 23, 67 (2 occurrences)
    # Plus 12 more IDs that appear twice each

    customers = [
        'Acme Corp', 'Blue Sky Ltd', 'Cascade Inc', 'Delta Solutions', 'Evergreen LLC',
        'Frontier Tech', 'Global Dynamics', 'Harbor Goods', 'Inland Traders', 'Jade Imports',
        'Keystone Partners', 'Landmark Corp', 'Maple Leaf Co', 'Nordic Supplies', 'Orbit Systems',
        'Pacific Rim Ltd', 'Quantum Labs', 'Riverdale Group', 'Sunrise Exports', 'Thunder Peak',
        'United Merchants', 'Valley Fresh', 'Westside Retail', 'XYZ Trading', 'Yellowstone Farms'
    ]

    dates = [
        '2025-01-03', '2025-01-07', '2025-01-10', '2025-01-14', '2025-01-17',
        '2025-01-21', '2025-01-24', '2025-01-28', '2025-02-03', '2025-02-06',
        '2025-02-10', '2025-02-13', '2025-02-17', '2025-02-20', '2025-02-24',
        '2025-02-27', '2025-03-03', '2025-03-06', '2025-03-10', '2025-03-13',
        '2025-03-17', '2025-03-20', '2025-03-24', '2025-03-27', '2025-03-31',
        '2025-04-03', '2025-04-07', '2025-04-10', '2025-04-14', '2025-04-17',
    ]

    amounts = [
        125.50, 340.00, 89.75, 1250.00, 567.30, 2340.50, 430.00, 760.00, 199.99, 880.00,
        1100.00, 245.00, 3200.00, 670.50, 490.00, 1550.00, 310.00, 780.00, 99.50, 2200.00,
        450.00, 1800.00, 630.00, 290.00, 1400.00, 520.00, 980.00, 175.00, 3500.00, 720.00,
    ]

    # All unique base IDs (105 unique + 14 that will be duplicated = more total occurrences)
    # We need exactly 119 rows total
    # Strategy: 14 IDs appear more than once:
    #   - ORD-2234: rows 5, 45, 89 (3x)
    #   - ORD-2891: rows 23, 67 (2x)
    #   - 12 more IDs each appearing exactly 2x
    # Total duplicate slots: 3 + 2 + 12*2 = 29 rows with duplicate IDs
    # Plus 119 - 29 = 90 unique IDs
    # Actually: 14 IDs appear more than once -> need to define carefully

    # Let's define: 14 "duplicate" IDs, one appears 3x, rest appear 2x = 1*3 + 13*2 = 29 slots
    # Unique IDs: 119 - 29 = 90

    # Generate unique base IDs
    all_order_ids = []

    # 14 duplicate IDs (these will appear multiple times)
    dup_ids = [
        'ORD-2234',  # appears 3x: rows 5, 45, 89
        'ORD-2891',  # appears 2x: rows 23, 67
        'ORD-3047',  # appears 2x
        'ORD-3158',  # appears 2x
        'ORD-3291',  # appears 2x
        'ORD-3402',  # appears 2x
        'ORD-3519',  # appears 2x
        'ORD-3634',  # appears 2x
        'ORD-3747',  # appears 2x
        'ORD-3856',  # appears 2x
        'ORD-3963',  # appears 2x
        'ORD-4078',  # appears 2x
        'ORD-4183',  # appears 2x
        'ORD-4294',  # appears 2x
    ]

    # Build the 119-element list of order IDs
    # Start with 119 unique IDs (numbered sequentially for non-duplicates)
    unique_counter = 2001
    row_ids = []
    for i in range(119):
        num = unique_counter + i
        row_ids.append(f'ORD-{num}')

    # Now place the duplicate IDs at specific positions (0-indexed for rows 2-120)
    # Row 5 -> index 3, row 45 -> index 43, row 89 -> index 87
    # Row 23 -> index 21, row 67 -> index 65

    # Positions for ORD-2234 (3x): rows 5, 45, 89 -> indices 3, 43, 87
    row_ids[3] = 'ORD-2234'
    row_ids[43] = 'ORD-2234'
    row_ids[87] = 'ORD-2234'

    # Positions for ORD-2891 (2x): rows 23, 67 -> indices 21, 65
    row_ids[21] = 'ORD-2891'
    row_ids[65] = 'ORD-2891'

    # Distribute remaining 12 duplicate IDs across the remaining rows
    # Place pairs spread out through the list
    remaining_dup_ids = dup_ids[2:]  # 12 IDs
    placement_pairs = [
        (1, 30),    # indices 1 and 30
        (5, 50),    # indices 5 and 50
        (8, 60),    # indices 8 and 60
        (11, 70),   # indices 11 and 70
        (14, 75),   # indices 14 and 75
        (17, 80),   # indices 17 and 80
        (20, 85),   # indices 20 and 85
        (25, 90),   # indices 25 and 90
        (28, 95),   # indices 28 and 95
        (33, 100),  # indices 33 and 100
        (37, 105),  # indices 37 and 105
        (40, 110),  # indices 40 and 110
    ]

    for dup_id, (pos1, pos2) in zip(remaining_dup_ids, placement_pairs):
        row_ids[pos1] = dup_id
        row_ids[pos2] = dup_id

    # Write data rows 2-120
    import random as rnd
    rnd.seed(42)

    for i, order_id in enumerate(row_ids):
        row = i + 2  # rows 2 through 120
        date = dates[i % len(dates)]
        customer = customers[i % len(customers)]
        amount = amounts[i % len(amounts)]

        ws.cell(row=row, column=1, value=order_id)
        ws.cell(row=row, column=2, value=date)
        ws.cell(row=row, column=3, value=customer)
        ws.cell(row=row, column=4, value=round(amount, 2))

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: OrderLog, Rows: 2-120 (119 records)')
    print(f'14 Order IDs appear more than once (no conditional formatting applied)')


create_initial()
