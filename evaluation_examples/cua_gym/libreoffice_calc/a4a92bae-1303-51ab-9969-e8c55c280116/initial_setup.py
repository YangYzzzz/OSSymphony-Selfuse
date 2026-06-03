"""
Initial Setup: Freeze the first row on the 'Data' sheet
Task ID: calc_ps_055
Domain: libreoffice_calc

Creates a spreadsheet with a 'Data' sheet containing 500 rows of data
with headers in row 1. No frozen panes.
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_055'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Headers in row 1
    headers = ['ID', 'Name', 'Value', 'Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic name pools
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
        'Carlos', 'Fatima', 'Oliver', 'Amara', 'Liam', 'Sofia', 'Wei',
        'Aisha', 'Noah', 'Isabella', 'Kenji', 'Zara', 'Ethan', 'Chloe',
        'Raj', 'Marta', 'Daniel', 'Yuki', 'Hassan', 'Natalia', 'Thomas',
        'Ling', 'Victor'
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim',
        'Rodriguez', 'Okafor', 'Mueller', 'Tanaka', 'Singh', 'Brown',
        'Garcia', 'Lee', 'Kowalski', 'Nguyen', 'Anderson', 'Yamamoto',
        'Patel', 'Fischer', 'Lopez', 'Wang', 'Martin', 'Ali', 'Taylor',
        'Santos', 'Berg', 'Nakamura', 'Ivanov', 'Costa'
    ]

    base_date = datetime(2023, 1, 1)

    # 500 rows of data
    for r in range(2, 502):
        row_id = 1000 + r - 1
        name = f'{random.choice(first_names)} {random.choice(last_names)}'
        value = round(random.uniform(10.0, 9999.99), 2)
        date = base_date + timedelta(days=random.randint(0, 900))

        ws.cell(row=r, column=1, value=row_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=value)
        ws.cell(row=r, column=4, value=date)
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14

    # Ensure NO freeze panes (the task is to add them)
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
