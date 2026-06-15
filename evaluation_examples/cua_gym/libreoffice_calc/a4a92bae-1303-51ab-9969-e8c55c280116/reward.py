"""
Reward Script: Freeze the first row on the 'Data' sheet
Task ID: calc_ps_055
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): freeze_panes is set to 'A2' (freezes row 1)
  Component 2 (0.3): 'Data' sheet exists with correct headers in row 1
                      AND freeze_panes is 'A2' (compound check anchored to change)
  Component 3 (0.2): Data integrity preserved (500 data rows still present)
                      AND freeze_panes is 'A2' (compound check anchored to change)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_055'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Data' sheet must exist
    if 'Data' not in wb.sheetnames:
        print(f"FAIL: 'Data' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Data']

    # Component 1: freeze_panes is set to 'A2' (0.5 points)
    # This is the core task requirement - freeze row 1.
    # freeze_panes='A2' means everything above row 2 (i.e., row 1) is frozen.
    try:
        fp = ws.freeze_panes
        if fp == 'A2':
            print(f"PASS: Component 1 - freeze_panes is 'A2' (row 1 frozen) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 - expected freeze_panes='A2', found: {repr(fp)}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Headers intact AND freeze_panes set (0.3 points)
    # Compound check: headers in row 1 are correct AND freeze is applied.
    # The headers alone are a precondition; the AND with freeze anchors this to the task change.
    try:
        fp = ws.freeze_panes
        h_a1 = ws['A1'].value
        h_b1 = ws['B1'].value
        h_c1 = ws['C1'].value
        h_d1 = ws['D1'].value
        headers_ok = (
            str(h_a1).strip() == 'ID' and
            str(h_b1).strip() == 'Name' and
            str(h_c1).strip() == 'Value' and
            str(h_d1).strip() == 'Date'
        )
        if headers_ok and fp == 'A2':
            print(f"PASS: Component 2 - Headers [ID, Name, Value, Date] intact AND freeze applied (0.3 pts)")
            total_score += 0.3
        else:
            if not headers_ok:
                print(f"FAIL: Component 2 - Headers mismatch: A1={h_a1}, B1={h_b1}, C1={h_c1}, D1={h_d1}")
            if fp != 'A2':
                print(f"FAIL: Component 2 - freeze_panes not 'A2': {repr(fp)}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Data integrity preserved AND freeze applied (0.2 points)
    # 500 data rows (rows 2-501) should still be present AND freeze is set.
    try:
        fp = ws.freeze_panes
        max_row = ws.max_row
        data_ok = max_row >= 501  # 1 header + 500 data rows
        if data_ok and fp == 'A2':
            print(f"PASS: Component 3 - {max_row} rows present (>=501) AND freeze applied (0.2 pts)")
            total_score += 0.2
        else:
            if not data_ok:
                print(f"FAIL: Component 3 - Expected >=501 rows, found: {max_row}")
            if fp != 'A2':
                print(f"FAIL: Component 3 - freeze_panes not 'A2': {repr(fp)}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
