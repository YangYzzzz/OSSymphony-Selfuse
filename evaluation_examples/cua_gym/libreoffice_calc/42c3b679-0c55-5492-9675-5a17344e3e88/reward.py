"""
Reward Script: Fill in Launch Site and Mission Control City for space missions
Task ID: osworld_multi_apps_conference_city_011
Domain: libreoffice_calc
Scoring:
  Component 1: Launch Site cells filled with accurate data for all 5 missions (0.1 each, max 0.5)
  Component 2: Mission Control City cells filled with accurate data for all 5 missions (0.1 each, max 0.5)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_011'

# Known acceptable keywords for each mission's Launch Site (case-insensitive)
# Based on task context: Crew Dragon/Artemis/Crew-3 -> Kennedy Space Center, FL
# Perseverance -> Cape Canaveral/Kennedy, Webb -> Kourou, French Guiana
LAUNCH_SITE_KEYWORDS = {
    'SpaceX Crew Dragon Demo-2':  ['kennedy', 'cape canaveral', 'ksc', 'launch complex'],
    'Artemis I':                  ['kennedy', 'cape canaveral', 'ksc', 'launch complex'],
    'Mars Perseverance Rover':    ['kennedy', 'cape canaveral', 'ksc', 'launch complex'],
    'James Webb Space Telescope': ['kourou', 'guiana', 'csg'],
    'Crew-3':                     ['kennedy', 'cape canaveral', 'ksc', 'launch complex'],
}

# Known acceptable keywords for each mission's Mission Control City (case-insensitive)
# Based on task context: Crew Dragon/Crew-3 -> Hawthorne, CA (SpaceX HQ)
# Artemis I -> Houston, TX (NASA JSC), Perseverance -> Pasadena, CA (JPL)
# Webb -> Baltimore, MD (STScI)
MISSION_CONTROL_KEYWORDS = {
    'SpaceX Crew Dragon Demo-2':  ['hawthorne', 'california', 'ca', 'spacex'],
    'Artemis I':                  ['houston', 'johnson', 'jsc', 'tx'],
    'Mars Perseverance Rover':    ['pasadena', 'jpl', 'jet propulsion', 'california', 'ca'],
    'James Webb Space Telescope': ['baltimore', 'stsci', 'space telescope', 'md'],
    'Crew-3':                     ['hawthorne', 'california', 'ca', 'spacex'],
}

MISSION_ORDER = [
    'SpaceX Crew Dragon Demo-2',
    'Artemis I',
    'Mars Perseverance Rover',
    'James Webb Space Telescope',
    'Crew-3',
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the SpaceMissions sheet (or active sheet)
    if 'SpaceMissions' in wb.sheetnames:
        ws = wb['SpaceMissions']
    else:
        ws = wb.active

    # Read the header row to identify column positions
    header_row = [ws.cell(row=1, column=c).value for c in range(1, 10)]
    try:
        col_mission = header_row.index('Mission Name') + 1
        col_launch = header_row.index('Launch Site') + 1
        col_control = header_row.index('Mission Control City') + 1
    except ValueError as e:
        print(f"CRITICAL: Missing expected header column: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Build a dict: mission_name -> (launch_site_value, mission_control_value)
    data = {}
    for row_idx in range(2, ws.max_row + 1):
        mission_name = ws.cell(row=row_idx, column=col_mission).value
        launch_site = ws.cell(row=row_idx, column=col_launch).value
        mission_control = ws.cell(row=row_idx, column=col_control).value
        if mission_name:
            data[mission_name.strip()] = (launch_site, mission_control)

    # Component 1: Launch Site filled with accurate data (0.1 per mission, up to 0.5)
    print("\n--- Component 1: Launch Site Verification (0.1 pts each, max 0.5) ---")
    for mission in MISSION_ORDER:
        try:
            if mission not in data:
                print(f"FAIL: Mission '{mission}' not found in spreadsheet")
                continue
            launch_val, _ = data[mission]
            if launch_val is None or str(launch_val).strip() == '':
                print(f"FAIL: Launch Site for '{mission}' is empty")
                continue
            launch_str = str(launch_val).strip().lower()
            keywords = LAUNCH_SITE_KEYWORDS.get(mission, [])
            if any(kw in launch_str for kw in keywords):
                print(f"PASS: Launch Site for '{mission}': '{launch_val}' (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Launch Site for '{mission}': '{launch_val}' — no expected keyword {keywords}")
        except Exception as e:
            print(f"ERROR: Checking launch site for '{mission}': {e}")

    # Component 2: Mission Control City filled with accurate data (0.1 per mission, up to 0.5)
    print("\n--- Component 2: Mission Control City Verification (0.1 pts each, max 0.5) ---")
    for mission in MISSION_ORDER:
        try:
            if mission not in data:
                print(f"FAIL: Mission '{mission}' not found in spreadsheet")
                continue
            _, control_val = data[mission]
            if control_val is None or str(control_val).strip() == '':
                print(f"FAIL: Mission Control City for '{mission}' is empty")
                continue
            control_str = str(control_val).strip().lower()
            keywords = MISSION_CONTROL_KEYWORDS.get(mission, [])
            if any(kw in control_str for kw in keywords):
                print(f"PASS: Mission Control City for '{mission}': '{control_val}' (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Mission Control City for '{mission}': '{control_val}' — no expected keyword {keywords}")
        except Exception as e:
            print(f"ERROR: Checking mission control city for '{mission}': {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.1f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
