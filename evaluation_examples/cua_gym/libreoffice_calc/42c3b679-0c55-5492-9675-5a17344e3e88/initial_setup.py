"""
Initial Setup: Space missions spreadsheet with empty Launch Site and Mission Control City columns.
Task ID: osworld_multi_apps_conference_city_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_conference_city_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SpaceMissions"

    # --- Headers ---
    headers = ['Mission Name', 'Year', 'Launch Site', 'Mission Control City']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Mission data (Launch Site and Mission Control City intentionally blank) ---
    missions = [
        ['SpaceX Crew Dragon Demo-2', 2020, None, None],
        ['Artemis I', 2022, None, None],
        ['Mars Perseverance Rover', 2021, None, None],
        ['James Webb Space Telescope', 2021, None, None],
        ['Crew-3', 2021, None, None],
    ]

    for r, row_data in enumerate(missions, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
