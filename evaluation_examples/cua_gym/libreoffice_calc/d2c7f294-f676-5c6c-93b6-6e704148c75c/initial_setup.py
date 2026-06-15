"""
Initial Setup: Create sales data spreadsheet for dynamic dashboard task
Task ID: calc_sales_069
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Data ---
    ws = wb.active
    ws.title = 'Data'

    # Headers
    ws.cell(row=1, column=1, value='Month')
    ws.cell(row=1, column=2, value='Revenue')

    # Monthly sales data (Jan-25 through Aug-25)
    data = [
        ['Jan-25', 85000],
        ['Feb-25', 92000],
        ['Mar-25', 78000],
        ['Apr-25', 105000],
        ['May-25', 98000],
        ['Jun-25', 115000],
        ['Jul-25', 96000],
        ['Aug-25', 110000],
    ]
    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
