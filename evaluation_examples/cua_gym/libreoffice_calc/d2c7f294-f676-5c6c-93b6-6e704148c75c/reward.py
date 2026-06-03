"""
Reward Script: Dynamic sales dashboard with OFFSET-based trailing formulas
Task ID: calc_sales_069
Domain: libreoffice_calc
Scoring:
  C1: Dashboard sheet exists (0.15)
  C2: Dashboard labels A1-A5 correct (0.20)
  C3: B1 COUNTA formula for total records (0.15)
  C4: B2 SUM+OFFSET formula for last 6 months total (0.20)
  C5: B3 AVERAGE+OFFSET formula for last 6 months avg (0.10)
  C6: B4 AVERAGE+OFFSET trailing 3 month avg (0.10)
  C7: B5 OFFSET formula for latest month (0.10)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_069'


def normalize_formula(f):
    """Normalize a formula for comparison: uppercase, strip spaces, unify separators."""
    if not isinstance(f, str):
        return ''
    f = f.upper().replace(' ', '')
    # Normalize Data. vs Data! (LibreOffice vs Excel cross-sheet ref)
    f = f.replace('DATA.', 'DATA!')
    return f


def formula_contains(actual, *keywords):
    """Check if normalized formula contains all keywords."""
    norm = normalize_formula(actual)
    return all(kw.upper().replace(' ', '') in norm for kw in keywords)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Dashboard sheet exists (0.15 points)
    try:
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        if 'dashboard' in sheet_names_lower:
            # Get the actual sheet name (case-insensitive match)
            dash_name = wb.sheetnames[sheet_names_lower.index('dashboard')]
            print(f"PASS: Component 1 — Dashboard sheet exists as '{dash_name}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No 'Dashboard' sheet found. Sheets: {wb.sheetnames}")
            # Without Dashboard sheet, no further checks possible
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb[dash_name]

    # Component 2: Dashboard labels A1-A5 correct (0.20 points)
    try:
        expected_labels = {
            'A1': 'total records',
            'A2': 'last 6 months total',
            'A3': 'last 6 months avg',
            'A4': 'trailing 3 month avg',
            'A5': 'latest month',
        }
        label_matches = 0
        for coord, expected in expected_labels.items():
            val = ws[coord].value
            if val and expected in str(val).lower():
                label_matches += 1
            else:
                print(f"  INFO: {coord} label: expected contains '{expected}', got {repr(val)}")

        if label_matches == 5:
            print(f"PASS: Component 2 — All 5 dashboard labels correct (0.20 pts)")
            total_score += 0.20
        elif label_matches >= 3:
            partial = round(0.20 * label_matches / 5, 2)
            print(f"PARTIAL: Component 2 — {label_matches}/5 labels correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {label_matches}/5 labels correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B1 formula - COUNTA for total records (0.15 points)
    try:
        b1_val = ws['B1'].value
        b1_norm = normalize_formula(b1_val)
        if b1_norm.startswith('=') and 'COUNTA' in b1_norm and 'DATA!' in b1_norm and 'B' in b1_norm:
            print(f"PASS: Component 3 — B1 has COUNTA formula: {b1_val} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — B1 expected COUNTA formula on Data!B, got: {repr(b1_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: B2 formula - SUM with OFFSET for last 6 months total (0.20 points)
    try:
        b2_val = ws['B2'].value
        b2_norm = normalize_formula(b2_val)
        if b2_norm.startswith('=') and 'SUM' in b2_norm and 'OFFSET' in b2_norm:
            # Check it references Data sheet and uses 6 for the count
            if 'DATA!' in b2_norm and '6' in b2_norm:
                print(f"PASS: Component 4 — B2 has SUM+OFFSET formula with 6-month range: {b2_val} (0.20 pts)")
                total_score += 0.20
            else:
                partial = 0.10
                print(f"PARTIAL: Component 4 — B2 has SUM+OFFSET but missing Data! ref or '6': {b2_val} ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 4 — B2 expected SUM+OFFSET formula, got: {repr(b2_val)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: B3 formula - AVERAGE with OFFSET for last 6 months avg (0.10 points)
    try:
        b3_val = ws['B3'].value
        b3_norm = normalize_formula(b3_val)
        if b3_norm.startswith('=') and 'AVERAGE' in b3_norm and 'OFFSET' in b3_norm:
            if 'DATA!' in b3_norm and '6' in b3_norm:
                print(f"PASS: Component 5 — B3 has AVERAGE+OFFSET formula with 6-month range: {b3_val} (0.10 pts)")
                total_score += 0.10
            else:
                partial = 0.05
                print(f"PARTIAL: Component 5 — B3 has AVERAGE+OFFSET but missing Data! ref or '6': {b3_val} ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 5 — B3 expected AVERAGE+OFFSET formula, got: {repr(b3_val)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: B4 formula - AVERAGE with OFFSET trailing 3 month avg (0.10 points)
    try:
        b4_val = ws['B4'].value
        b4_norm = normalize_formula(b4_val)
        if b4_norm.startswith('=') and 'AVERAGE' in b4_norm and 'OFFSET' in b4_norm:
            if 'DATA!' in b4_norm and '3' in b4_norm:
                print(f"PASS: Component 6 — B4 has AVERAGE+OFFSET formula with 3-month range: {b4_val} (0.10 pts)")
                total_score += 0.10
            else:
                partial = 0.05
                print(f"PARTIAL: Component 6 — B4 has AVERAGE+OFFSET but missing Data! ref or '3': {b4_val} ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 6 — B4 expected AVERAGE+OFFSET formula, got: {repr(b4_val)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: B5 formula - OFFSET for latest month (0.10 points)
    try:
        b5_val = ws['B5'].value
        b5_norm = normalize_formula(b5_val)
        if b5_norm.startswith('=') and 'OFFSET' in b5_norm and 'DATA!' in b5_norm:
            print(f"PASS: Component 7 — B5 has OFFSET formula referencing Data: {b5_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 — B5 expected OFFSET formula referencing Data, got: {repr(b5_val)}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
