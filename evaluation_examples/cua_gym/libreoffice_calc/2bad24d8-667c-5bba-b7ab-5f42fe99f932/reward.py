"""
Reward Script: Build a sales KPI dashboard with large formatted metric cells and a pipeline funnel chart.
Task ID: calc_gpm_016
Domain: libreoffice_calc
Scoring:
  Component 1: Title merge + formatting (A1:J1) — 0.15
  Component 2: KPI cell merges (rows 3-5) — 0.15
  Component 3: KPI value formatting (row 4: 20pt bold) — 0.15
  Component 4: KPI fill colors (accent tints) — 0.15
  Component 5: Pipeline header formatting (row 8) — 0.15
  Component 6: Thick borders on KPI boxes — 0.10
  Component 7: Bar chart with title 'Sales Pipeline' — 0.15
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_016'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'SalesDash' not in wb.sheetnames:
        print("FAIL: Sheet 'SalesDash' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SalesDash']
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]

    # Component 1: Title merge + formatting (0.15 points)
    # A1:J1 should be merged, value = 'Sales Dashboard - March 2026', 18pt bold, white font on dark blue fill
    try:
        title_merged = any('A1' in r and 'J1' in r for r in merged_ranges)
        a1 = ws['A1']
        title_value = str(a1.value).strip() if a1.value else ''
        title_bold = a1.font.bold is True
        title_size = a1.font.size is not None and a1.font.size >= 16

        # Check dark blue fill (003366 or similar)
        try:
            bg_rgb = a1.fill.fgColor.rgb
            has_dark_fill = bg_rgb is not None and a1.fill.patternType == 'solid'
        except:
            has_dark_fill = False

        if title_merged and 'Sales Dashboard' in title_value and title_bold and title_size and has_dark_fill:
            print(f"PASS: Component 1 — Title merged A1:J1, bold {a1.font.size}pt, dark fill (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — merged={title_merged}, value='{title_value}', bold={title_bold}, size_ok={title_size}, fill={has_dark_fill}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: KPI cell merges (0.15 points)
    # Rows 3-5 should have merged pairs: A3:B3, C3:D3, E3:F3, G3:H3, and similarly for rows 4 and 5
    try:
        # Check that at least 8 of the expected 12 merge ranges exist (some tolerance)
        expected_merges = [
            'A3:B3', 'C3:D3', 'E3:F3', 'G3:H3',
            'A4:B4', 'C4:D4', 'E4:F4', 'G4:H4',
            'A5:B5', 'C5:D5', 'E5:F5', 'G5:H5',
        ]
        found_count = 0
        for em in expected_merges:
            if em in merged_ranges:
                found_count += 1

        if found_count >= 8:
            print(f"PASS: Component 2 — {found_count}/12 KPI merges found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Only {found_count}/12 KPI merges found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: KPI value formatting — row 4 cells should be 20pt bold (0.15 points)
    # A4, C4, E4, G4 should have 20pt bold fonts
    try:
        kpi_value_cells = ['A4', 'C4', 'E4', 'G4']
        formatted_count = 0
        for coord in kpi_value_cells:
            cell = ws[coord]
            if cell.font.bold is True and cell.font.size is not None and cell.font.size >= 18:
                formatted_count += 1
            else:
                print(f"  DEBUG: {coord} font: bold={cell.font.bold}, size={cell.font.size}")

        if formatted_count >= 3:
            print(f"PASS: Component 3 — {formatted_count}/4 KPI values are large bold (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Only {formatted_count}/4 KPI values formatted correctly")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: KPI fill colors — each KPI box should have a distinct colored background (0.15 points)
    # Revenue=green tint, Deals=blue tint, Avg Deal=neutral, Win Rate=orange tint
    try:
        kpi_cells = ['A3', 'C3', 'E3', 'G3']
        colored_count = 0
        for coord in kpi_cells:
            cell = ws[coord]
            try:
                bg = cell.fill.fgColor.rgb
                fill_type = cell.fill.patternType
                # Must have a solid fill that is NOT default (00000000)
                if fill_type == 'solid' and bg is not None and bg != '00000000':
                    colored_count += 1
                else:
                    print(f"  DEBUG: {coord} fill: type={fill_type}, bg={bg}")
            except:
                print(f"  DEBUG: {coord} no fill color")

        if colored_count >= 3:
            print(f"PASS: Component 4 — {colored_count}/4 KPI boxes have accent fills (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Only {colored_count}/4 KPI boxes have colored fills")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Pipeline header formatting — row 8 bold, dark blue fill, white text (0.15 points)
    try:
        header_cells = ['A8', 'B8', 'C8']
        header_ok = 0
        for coord in header_cells:
            cell = ws[coord]
            is_bold = cell.font.bold is True
            try:
                bg = cell.fill.fgColor.rgb
                has_fill = cell.fill.patternType == 'solid' and bg is not None and bg != '00000000'
            except:
                has_fill = False
            try:
                fc = cell.font.color.rgb
                # White text: FFFFFF or 00FFFFFF
                has_white = fc is not None and 'FFFFFF' in fc
            except:
                has_white = False

            if is_bold and has_fill and has_white:
                header_ok += 1
            else:
                print(f"  DEBUG: {coord} bold={is_bold}, fill={has_fill}, white_text={has_white}")

        if header_ok >= 2:
            print(f"PASS: Component 5 — {header_ok}/3 pipeline headers formatted (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 5 — Only {header_ok}/3 pipeline headers formatted correctly")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Thick borders on KPI boxes (0.10 points)
    # Check that KPI box corners have thick border style
    try:
        border_check_cells = ['A3', 'A5', 'C3', 'E3', 'G3']
        border_count = 0
        for coord in border_check_cells:
            cell = ws[coord]
            has_thick = False
            for side_name in ['left', 'right', 'top', 'bottom']:
                side = getattr(cell.border, side_name)
                if side and side.style == 'thick':
                    has_thick = True
                    break
            if has_thick:
                border_count += 1
            else:
                print(f"  DEBUG: {coord} no thick border found")

        if border_count >= 3:
            print(f"PASS: Component 6 — {border_count}/5 KPI cells have thick borders (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 6 — Only {border_count}/5 KPI cells have thick borders")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Bar chart with title 'Sales Pipeline' (0.15 points)
    try:
        charts = ws._charts
        if len(charts) >= 1:
            found_pipeline_chart = False
            for ch in charts:
                chart_type = ch.type  # 'bar' for horizontal bar chart
                # Extract title text
                title_text = ''
                try:
                    for p in ch.title.tx.rich.paragraphs:
                        for r in p.r:
                            title_text += r.t
                except:
                    pass

                if chart_type == 'bar' and 'Pipeline' in title_text:
                    found_pipeline_chart = True
                    break

            if found_pipeline_chart:
                print(f"PASS: Component 7 — Bar chart titled 'Sales Pipeline' found (0.15 pts)")
                total_score += 0.15
            else:
                # Partial credit: chart exists but wrong type or title
                chart_types = [c.type for c in charts]
                print(f"FAIL: Component 7 — Charts found ({chart_types}) but no bar chart titled 'Sales Pipeline'")
        else:
            print(f"FAIL: Component 7 — No charts found in sheet")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
