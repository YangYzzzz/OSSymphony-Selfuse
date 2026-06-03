"""
Initial Setup: Build a sales KPI dashboard with large formatted metric cells and a pipeline funnel chart.
Task ID: calc_gpm_016
Domain: libreoffice_calc

Creates initial file with raw, unformatted data that the agent will transform into a dashboard.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SalesDash'

    # --- Row 1: Title text (unformatted, not merged) ---
    ws['A1'] = 'Sales Dashboard - March 2026'

    # --- Rows 3-5: KPI raw data (unformatted, not merged) ---
    # Revenue KPI
    ws['A3'] = 'Monthly Revenue'
    ws['A4'] = '$1,245,000'
    ws['A5'] = 'vs Target: +12%'

    # Deals Won KPI
    ws['C3'] = 'Deals Closed'
    ws['C4'] = '23'
    ws['C5'] = 'vs Last Month: +5'

    # Avg Deal Size KPI
    ws['E3'] = 'Avg Deal'
    ws['E4'] = '$54,130'
    ws['E5'] = 'vs Target: -3%'

    # Win Rate KPI
    ws['G3'] = 'Win Rate'
    ws['G4'] = '34%'
    ws['G5'] = 'Target: 40%'

    # --- Pipeline data (rows 8-13, unformatted headers) ---
    ws['A8'] = 'Stage'
    ws['B8'] = 'Deals'
    ws['C8'] = 'Value'

    pipeline_data = [
        ['Prospect', 45, 2250000],
        ['Qualified', 28, 1680000],
        ['Proposal', 15, 1050000],
        ['Negotiation', 8, 640000],
        ['Closed Won', 23, 1245000],
    ]
    for r, row_data in enumerate(pipeline_data, 9):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
