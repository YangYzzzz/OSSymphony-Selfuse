"""
Initial Setup: Customer Satisfaction Survey Results Analyzer
Task ID: calc_grs_019
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Survey Data ---
    ws = wb.active
    ws.title = "Survey Data"

    headers = [
        "Response ID", "Date", "Customer Segment", "NPS Score",
        "Satisfaction Rating", "Response Time Rating",
        "Product Quality Rating", "Comments Category"
    ]
    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 22
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 20

    # Generate 40 realistic survey responses across 3 months
    segments = ["Enterprise", "SMB", "Consumer"]
    comment_categories = [
        "Praise", "Feature Request", "Bug Report",
        "Support Issue", "Pricing Concern", "General Feedback"
    ]

    first_names = [
        "Sarah", "Marcus", "Elena", "James", "Priya", "David", "Mei",
        "Carlos", "Fatima", "Ryan", "Aisha", "Thomas", "Yuki", "Michael",
        "Sofia", "Andre", "Rachel", "Omar", "Lily", "Nathan",
        "Grace", "Viktor", "Hannah", "Raj", "Emma", "Kenji", "Olivia",
        "Samuel", "Anya", "Derek", "Zara", "Felix", "Chloe", "Ibrahim",
        "Mia", "Ethan", "Nadia", "Lucas", "Amira", "Jordan"
    ]

    random.seed(42)  # reproducibility

    # Dates spanning Jan-Mar 2025
    base_dates = []
    for month in [1, 2, 3]:
        for _ in range(14 if month <= 2 else 12):
            day = random.randint(1, 28)
            base_dates.append(datetime(2025, month, day))
    base_dates.sort()
    base_dates = base_dates[:40]

    data_rows = []
    for i in range(40):
        resp_id = f"SR-2025-{i+1:04d}"
        date = base_dates[i].strftime("%Y-%m-%d")
        segment = random.choice(segments)

        # NPS scores vary by segment to create interesting patterns
        if segment == "Enterprise":
            nps = random.choices(range(0, 11), weights=[1,1,1,1,1,2,3,4,5,6,7])[0]
        elif segment == "SMB":
            nps = random.choices(range(0, 11), weights=[1,2,2,2,3,3,4,4,4,5,5])[0]
        else:
            nps = random.choices(range(0, 11), weights=[2,2,3,3,3,4,4,3,3,3,3])[0]

        satisfaction = min(5, max(1, int(nps / 2.5) + random.randint(0, 1)))
        response_time = random.randint(1, 5)
        product_quality = min(5, max(1, int(nps / 2.2) + random.randint(-1, 1)))
        comment_cat = random.choice(comment_categories)

        data_rows.append([
            resp_id, date, segment, nps,
            satisfaction, response_time, product_quality, comment_cat
        ])

    data_align = Alignment(horizontal="center", vertical="center")
    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = data_align
            cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Sheet 2: Analysis (empty - task requires building this) ---
    ws2 = wb.create_sheet("Analysis")
    # Add a simple header to indicate this sheet is for analysis
    ws2["A1"] = "Analysis"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True)
    ws2["A3"] = "Complete the analysis below based on Survey Data"
    ws2["A3"].font = Font(name="Calibri", size=11, italic=True, color="808080")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
