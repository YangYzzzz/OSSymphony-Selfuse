"""
Reward Script: Customer Satisfaction Survey Results Analyzer
Task ID: calc_grs_019
Domain: libreoffice_calc
Scoring:
  Component 1: NPS calculation section with formulas (0.20)
  Component 2: Average scores by dimension with AVERAGE formulas (0.15)
  Component 3: Cross-tabulation by segment with AVERAGEIF formulas (0.20)
  Component 4: Diverging color scale conditional formatting on Survey Data (0.15)
  Component 5: Radar chart present on Analysis sheet (0.15)
  Component 6: Monthly NPS trend section + line chart (0.15)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_019'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Survey Data sheet must exist with data
    if 'Survey Data' not in wb.sheetnames:
        print("CRITICAL: 'Survey Data' sheet missing")
        print("REWARD: 0.0")
        return 0.0

    ws_survey = wb['Survey Data']

    # Precondition: Analysis sheet must exist
    analysis_name = None
    for sn in wb.sheetnames:
        if sn.lower() == 'analysis':
            analysis_name = sn
            break
    if analysis_name is None:
        print("CRITICAL: No 'Analysis' sheet found")
        print("REWARD: 0.0")
        return 0.0

    ws_analysis = wb[analysis_name]

    # =====================================================================
    # Component 1: NPS Calculation Section (0.20 points)
    # Golden has: COUNTA for total, COUNTIF for promoters/detractors,
    # NPS formula = (promoters/total - detractors/total)*100
    # Initial Analysis sheet has none of these formulas.
    # =====================================================================
    try:
        nps_score_comp = 0.0
        # Scan all cells in Analysis for NPS-related formulas
        has_countif = False
        has_nps_formula = False
        has_promoter_count = False
        has_detractor_count = False

        for r in range(1, ws_analysis.max_row + 1):
            for c in range(1, ws_analysis.max_column + 1):
                val = ws_analysis.cell(row=r, column=c).value
                if isinstance(val, str):
                    val_upper = val.upper().replace(" ", "")
                    # Check for COUNTIF referencing NPS Score column (D)
                    if 'COUNTIF' in val_upper and 'SURVEYDATA' in val_upper.replace("'", "").replace(" ", ""):
                        has_countif = True
                        if '>=9' in val_upper or '>="&9' in val_upper or '">=9"' in val_upper or '">="&9' in val_upper:
                            has_promoter_count = True
                        if '<7' in val_upper or '"<"&7' in val_upper or '"<7"' in val_upper:
                            has_detractor_count = True
                    # Check for NPS calculation (subtraction of percentages * 100)
                    if '*100' in val_upper and ('B5' in val_upper or 'B7' in val_upper or 'PROMOTER' in val_upper.upper()):
                        has_nps_formula = True
                    # Alternative: any formula that multiplies by 100 and involves division
                    if '*100' in val_upper and '/' in val_upper:
                        has_nps_formula = True

        if has_countif and has_promoter_count:
            nps_score_comp += 0.08
            print("PASS: Component 1a - COUNTIF formula for promoters found")
        else:
            print(f"FAIL: Component 1a - Missing COUNTIF for promoters (has_countif={has_countif}, has_promoter={has_promoter_count})")

        if has_detractor_count:
            nps_score_comp += 0.06
            print("PASS: Component 1b - COUNTIF formula for detractors found")
        else:
            print(f"FAIL: Component 1b - Missing COUNTIF for detractors")

        if has_nps_formula:
            nps_score_comp += 0.06
            print("PASS: Component 1c - NPS calculation formula found")
        else:
            print(f"FAIL: Component 1c - Missing NPS formula (% Promoters - % Detractors)*100")

        total_score += nps_score_comp
        print(f"  Component 1 subtotal: {nps_score_comp}/0.20")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =====================================================================
    # Component 2: Average Scores by Dimension (0.15 points)
    # Golden has AVERAGE formulas referencing Survey Data columns E, F, G
    # for Satisfaction, Response Time, Product Quality ratings.
    # Initial has none of these.
    # =====================================================================
    try:
        avg_comp = 0.0
        avg_cols_found = set()

        for r in range(1, ws_analysis.max_row + 1):
            for c in range(1, ws_analysis.max_column + 1):
                val = ws_analysis.cell(row=r, column=c).value
                if isinstance(val, str):
                    val_clean = val.upper().replace(" ", "").replace("'", "")
                    # AVERAGE formula referencing Survey Data columns
                    if 'AVERAGE(' in val_clean and 'SURVEYDATA' in val_clean:
                        if '!E' in val.replace(" ", "").replace("'", ""):
                            avg_cols_found.add('E')
                        if '!F' in val.replace(" ", "").replace("'", ""):
                            avg_cols_found.add('F')
                        if '!G' in val.replace(" ", "").replace("'", ""):
                            avg_cols_found.add('G')

        if 'E' in avg_cols_found:
            avg_comp += 0.05
            print("PASS: Component 2a - AVERAGE for Satisfaction Rating (col E)")
        else:
            print("FAIL: Component 2a - Missing AVERAGE for Satisfaction Rating")

        if 'F' in avg_cols_found:
            avg_comp += 0.05
            print("PASS: Component 2b - AVERAGE for Response Time Rating (col F)")
        else:
            print("FAIL: Component 2b - Missing AVERAGE for Response Time Rating")

        if 'G' in avg_cols_found:
            avg_comp += 0.05
            print("PASS: Component 2c - AVERAGE for Product Quality Rating (col G)")
        else:
            print("FAIL: Component 2c - Missing AVERAGE for Product Quality Rating")

        total_score += avg_comp
        print(f"  Component 2 subtotal: {avg_comp}/0.15")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =====================================================================
    # Component 3: Cross-tabulation by Segment with AVERAGEIF (0.20 points)
    # Golden has AVERAGEIF formulas for Enterprise, SMB, Consumer segments
    # across multiple rating dimensions. Initial has none.
    # =====================================================================
    try:
        seg_comp = 0.0
        segments_found = set()

        for r in range(1, ws_analysis.max_row + 1):
            for c in range(1, ws_analysis.max_column + 1):
                val = ws_analysis.cell(row=r, column=c).value
                if isinstance(val, str):
                    val_upper = val.upper().replace(" ", "")
                    if 'AVERAGEIF' in val_upper:
                        if 'ENTERPRISE' in val_upper:
                            segments_found.add('Enterprise')
                        if 'SMB' in val_upper:
                            segments_found.add('SMB')
                        if 'CONSUMER' in val_upper:
                            segments_found.add('Consumer')

        for seg in ['Enterprise', 'SMB', 'Consumer']:
            if seg in segments_found:
                seg_comp += 0.067
                print(f"PASS: Component 3 - AVERAGEIF for '{seg}' segment found")
            else:
                print(f"FAIL: Component 3 - Missing AVERAGEIF for '{seg}' segment")

        # Round to avoid floating point issues
        seg_comp = round(min(seg_comp, 0.20), 2)
        total_score += seg_comp
        print(f"  Component 3 subtotal: {seg_comp}/0.20")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =====================================================================
    # Component 4: Diverging Color Scale Conditional Formatting (0.15 points)
    # Golden has 4 color scale rules on Survey Data (D2:D41, E2:E41,
    # F2:F41, G2:G41). Initial has 0 rules.
    # =====================================================================
    try:
        cf_comp = 0.0
        color_scale_count = 0

        cf_rules = ws_survey.conditional_formatting
        for cf in cf_rules:
            for rule in cf.rules:
                if rule.type == 'colorScale':
                    color_scale_count += 1

        if color_scale_count >= 4:
            cf_comp = 0.15
            print(f"PASS: Component 4 - {color_scale_count} color scale rules found (need >= 4)")
        elif color_scale_count >= 2:
            cf_comp = 0.08
            print(f"PARTIAL: Component 4 - {color_scale_count} color scale rules found (need >= 4)")
        elif color_scale_count >= 1:
            cf_comp = 0.04
            print(f"PARTIAL: Component 4 - {color_scale_count} color scale rule found (need >= 4)")
        else:
            print(f"FAIL: Component 4 - No color scale conditional formatting found on Survey Data")

        total_score += cf_comp
        print(f"  Component 4 subtotal: {cf_comp}/0.15")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =====================================================================
    # Component 5: Radar/Spider Chart (0.15 points)
    # Golden has a RadarChart on Analysis sheet. Initial has 0 charts.
    # =====================================================================
    try:
        chart_comp = 0.0
        charts = ws_analysis._charts
        radar_found = False
        any_chart_found = len(charts) > 0

        for ch in charts:
            class_name = ch.__class__.__name__
            if 'Radar' in class_name:
                radar_found = True
                break

        if radar_found:
            chart_comp = 0.15
            print("PASS: Component 5 - Radar chart found on Analysis sheet")
        elif any_chart_found:
            chart_comp = 0.05
            print(f"PARTIAL: Component 5 - Charts found but no radar chart (classes: {[c.__class__.__name__ for c in charts]})")
        else:
            print("FAIL: Component 5 - No charts found on Analysis sheet")

        total_score += chart_comp
        print(f"  Component 5 subtotal: {chart_comp}/0.15")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # =====================================================================
    # Component 6: Monthly NPS Trend Section + Line Chart (0.15 points)
    # Golden has monthly trend data with COUNTIFS by month and a LineChart.
    # Initial has neither.
    # =====================================================================
    try:
        trend_comp = 0.0

        # Check for monthly COUNTIFS formulas (month-based filtering)
        has_monthly_countifs = False
        for r in range(1, ws_analysis.max_row + 1):
            for c in range(1, ws_analysis.max_column + 1):
                val = ws_analysis.cell(row=r, column=c).value
                if isinstance(val, str):
                    val_upper = val.upper().replace(" ", "")
                    # Looking for COUNTIFS with date pattern like "2025-01-*"
                    if 'COUNTIFS' in val_upper and ('2025-01' in val or '2025-02' in val or '2025-03' in val):
                        has_monthly_countifs = True
                        break
            if has_monthly_countifs:
                break

        if has_monthly_countifs:
            trend_comp += 0.08
            print("PASS: Component 6a - Monthly COUNTIFS formulas found for NPS trend")
        else:
            print("FAIL: Component 6a - No monthly COUNTIFS formulas found")

        # Check for line chart
        line_chart_found = False
        for ch in ws_analysis._charts:
            class_name = ch.__class__.__name__
            if 'Line' in class_name:
                line_chart_found = True
                break

        if line_chart_found:
            trend_comp += 0.07
            print("PASS: Component 6b - Line chart found on Analysis sheet")
        else:
            print("FAIL: Component 6b - No line chart found on Analysis sheet")

        total_score += trend_comp
        print(f"  Component 6 subtotal: {trend_comp}/0.15")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Final score
    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state('libreoffice_calc')

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
