"""
Initial Setup: IFERROR+VLOOKUP product name lookup task
Task ID: calc_fma_iferror_vlookup_006
Domain: libreoffice_calc
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_iferror_vlookup_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Orders ---
    ws_orders = wb.active
    ws_orders.title = 'Orders'

    # Headers
    ws_orders.cell(row=1, column=1, value='Product Code')
    ws_orders.cell(row=1, column=2, value='Product Name')

    # Product codes in column A rows 2-20 (exact order from context)
    product_codes = [
        'PRD-001', 'PRD-003', 'PRD-007', 'PRD-002', 'PRD-099',
        'PRD-005', 'PRD-004', 'PRD-011', 'PRD-088', 'PRD-006',
        'PRD-001', 'PRD-003', 'PRD-200', 'PRD-008', 'PRD-002',
        'PRD-009', 'PRD-077', 'PRD-010', 'PRD-005'
    ]

    for i, code in enumerate(product_codes, 2):
        ws_orders.cell(row=i, column=1, value=code)
        # Column B is intentionally EMPTY — the task is to fill it with IFERROR/VLOOKUP formulas

    # Set column widths for readability
    ws_orders.column_dimensions['A'].width = 15
    ws_orders.column_dimensions['B'].width = 20

    # --- Sheet 2: Catalog ---
    ws_catalog = wb.create_sheet('Catalog')

    # Headers
    ws_catalog.cell(row=1, column=1, value='Product Code')
    ws_catalog.cell(row=1, column=2, value='Product Name')

    # Product catalog data (PRD-001 to PRD-011)
    catalog_data = [
        ('PRD-001', 'Pen'),
        ('PRD-002', 'Notebook'),
        ('PRD-003', 'Stapler'),
        ('PRD-004', 'Tape'),
        ('PRD-005', 'Scissors'),
        ('PRD-006', 'Ruler'),
        ('PRD-007', 'Eraser'),
        ('PRD-008', 'Marker'),
        ('PRD-009', 'Folder'),
        ('PRD-010', 'Binder'),
        ('PRD-011', 'Calculator'),
    ]

    for i, (code, name) in enumerate(catalog_data, 2):
        ws_catalog.cell(row=i, column=1, value=code)
        ws_catalog.cell(row=i, column=2, value=name)

    # Set column widths for readability
    ws_catalog.column_dimensions['A'].width = 15
    ws_catalog.column_dimensions['B'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Orders sheet: rows 2-20 with product codes, column B empty')
    print(f'  Catalog sheet: 11 product entries (PRD-001 to PRD-011)')


create_initial()
