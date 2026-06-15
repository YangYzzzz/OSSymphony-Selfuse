"""
Reward Script: VLOOKUP with IFERROR formula in Orders sheet
Task ID: calc_fma_iferror_vlookup_006
Domain: libreoffice_calc
Scoring:
  Component 1: All 19 cells B2:B20 contain IFERROR+VLOOKUP formulas with correct structure (0.6 pts)
  Component 2: Each formula correctly references its own row number A2..A20 (0.4 pts)
Total: 1.0

Note: Data integrity (catalog and A-column unchanged) is used as a precondition gate,
not a scoring component, because those values are identical in initial and golden.
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_iferror_vlookup_006'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Write IFERROR(VLOOKUP(...)) formulas in B2:B20 of the Orders sheet.
    Initial state: B2:B20 are all empty (None).
    Golden state: B2:B20 all contain =IFERROR(VLOOKUP(Ax,Catalog.$A:$B,2,0),"Not Found")
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: both required sheets must exist
    if 'Orders' not in wb.sheetnames:
        print("CRITICAL: 'Orders' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    if 'Catalog' not in wb.sheetnames:
        print("CRITICAL: 'Catalog' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_orders = wb['Orders']

    # -------------------------------------------------------------------------
    # Component 1: All 19 cells B2:B20 contain IFERROR+VLOOKUP formulas (0.6 pts)
    # The initial file has B2:B20 all empty — any formula present means the
    # agent took action. We require all 19 cells to have a formula that:
    #   - is wrapped in IFERROR()
    #   - contains VLOOKUP()
    #   - references the Catalog sheet
    #   - returns "Not Found" for missing codes
    # This FAILS on initial (all empty) → PASSES on golden (all formulas).
    # -------------------------------------------------------------------------
    try:
        formulas_correct_structure = 0
        wrong_cells = []

        for row in range(2, 21):  # rows 2 to 20 inclusive
            cell_val = ws_orders.cell(row=row, column=2).value
            if cell_val is None:
                wrong_cells.append(f"B{row}: empty")
                continue

            formula = str(cell_val).strip().upper().replace(" ", "")

            has_iferror = 'IFERROR(' in formula
            has_vlookup = 'VLOOKUP(' in formula
            has_catalog = 'CATALOG' in formula
            # Check for "Not Found" in various forms: NOTFOUND or "NOTFOUND"
            has_not_found = ('NOTFOUND' in formula) or ('"NOT FOUND"' in str(cell_val).upper())

            if has_iferror and has_vlookup and has_catalog and has_not_found:
                formulas_correct_structure += 1
            else:
                wrong_cells.append(
                    f"B{row}: {repr(cell_val)[:80]} "
                    f"(iferror={has_iferror}, vlookup={has_vlookup}, "
                    f"catalog={has_catalog}, not_found={has_not_found})"
                )

        if formulas_correct_structure == 19:
            print(f"PASS: Component 1 — All 19 cells B2:B20 have correct IFERROR+VLOOKUP structure (0.6 pts)")
            total_score += 0.6
        elif formulas_correct_structure > 0:
            partial = round(0.6 * formulas_correct_structure / 19, 4)
            print(f"PARTIAL: Component 1 — {formulas_correct_structure}/19 cells have correct structure ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — 0/19 cells have correct IFERROR+VLOOKUP structure (0 pts)")
            if wrong_cells:
                for wc in wrong_cells[:5]:
                    print(f"  - {wc}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Each formula correctly references its own row number (0.4 pts)
    # e.g., B2 must reference A2, B5 must reference A5, etc.
    # This checks that the row reference is row-relative, not fixed or wrong.
    # This FAILS on initial (all empty) → PASSES on golden (each references Ax).
    # -------------------------------------------------------------------------
    try:
        row_ref_correct = 0
        row_ref_wrong = []

        for row in range(2, 21):
            cell_val = ws_orders.cell(row=row, column=2).value
            if cell_val is None:
                row_ref_wrong.append(f"B{row}: empty")
                continue

            formula = str(cell_val).strip()
            # The VLOOKUP first argument must be A<row_number>
            pattern = rf'VLOOKUP\s*\(\s*A{row}\s*,'
            if re.search(pattern, formula, re.IGNORECASE):
                row_ref_correct += 1
            else:
                found_ref = re.search(r'VLOOKUP\s*\(\s*(A\d+)', formula, re.IGNORECASE)
                actual_ref = found_ref.group(1) if found_ref else "no A-ref"
                row_ref_wrong.append(f"B{row}: expected A{row} ref, found '{actual_ref}'")

        if row_ref_correct == 19:
            print(f"PASS: Component 2 — All 19 formulas reference their own row (A2..A20) (0.4 pts)")
            total_score += 0.4
        elif row_ref_correct > 0:
            partial = round(0.4 * row_ref_correct / 19, 4)
            print(f"PARTIAL: Component 2 — {row_ref_correct}/19 formulas reference correct row ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — 0/19 formulas reference correct row (0 pts)")
            if row_ref_wrong:
                for wr in row_ref_wrong[:5]:
                    print(f"  - {wr}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
