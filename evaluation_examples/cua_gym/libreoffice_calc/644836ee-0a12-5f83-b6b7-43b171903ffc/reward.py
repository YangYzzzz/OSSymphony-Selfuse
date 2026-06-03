"""
Reward Script: Format weekly project tracker with merged title, orange milestone rows
Task ID: calc_gsd_020
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): Title row A1 is 16pt bold
  Component 2 (0.4): Milestone rows (7,15,24,33,44) cols A-G have orange (#FF6600) background
  Component 3 (0.4): Milestone rows (7,15,24,33,44) cols A-G have bold white text
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_020'
MILESTONE_ROWS = [7, 15, 24, 33, 44]
COLS = range(1, 8)  # columns A(1) through G(7)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Timeline' sheet must exist
    if 'Timeline' not in wb.sheetnames:
        print("CRITICAL: 'Timeline' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Timeline']

    # Component 1: Title row A1 is 16pt bold (0.2 points)
    try:
        a1 = ws['A1']
        is_bold = a1.font.bold == True
        is_16pt = False
        if a1.font.size is not None:
            is_16pt = abs(float(a1.font.size) - 16.0) < 0.5

        if is_bold and is_16pt:
            print(f"PASS: Component 1 — A1 is {a1.font.size}pt bold (0.2 pts)")
            total_score += 0.2
        elif is_bold:
            print(f"PARTIAL: Component 1 — A1 is bold but size={a1.font.size} (not 16pt) (0.1 pts)")
            total_score += 0.1
        elif is_16pt:
            print(f"PARTIAL: Component 1 — A1 is 16pt but not bold (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 1 — A1 bold={a1.font.bold}, size={a1.font.size}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Orange background (#FF6600) on milestone rows A-G (0.4 points)
    # Each milestone row checked across 7 cols = 35 cells total
    # Score proportionally: 0.4 * (passing_cells / 35)
    try:
        orange_pass = 0
        orange_total = len(MILESTONE_ROWS) * len(list(COLS))  # 35
        for r in MILESTONE_ROWS:
            for c in COLS:
                cell = ws.cell(row=r, column=c)
                try:
                    fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                    # Accept FFFF6600 (exact match for #FF6600 with full alpha)
                    if fill_rgb == 'FFFF6600' and cell.fill.fill_type == 'solid':
                        orange_pass += 1
                except Exception:
                    pass

        if orange_pass == orange_total:
            print(f"PASS: Component 2 — All {orange_total} milestone cells have orange background (0.4 pts)")
            total_score += 0.4
        elif orange_pass > 0:
            partial = round(0.4 * (orange_pass / orange_total), 2)
            print(f"PARTIAL: Component 2 — {orange_pass}/{orange_total} milestone cells have orange bg ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No milestone cells have orange background")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Bold white text on milestone rows A-G (0.4 points)
    # Check bold=True and font color is white (FFFFFF)
    # Score proportionally: 0.4 * (passing_cells / 35)
    try:
        text_pass = 0
        text_total = len(MILESTONE_ROWS) * len(list(COLS))  # 35
        for r in MILESTONE_ROWS:
            for c in COLS:
                cell = ws.cell(row=r, column=c)
                try:
                    is_bold = cell.font.bold == True
                    font_color = None
                    if cell.font.color and cell.font.color.rgb:
                        font_color = cell.font.color.rgb
                    # Accept white: 00FFFFFF or FFFFFFFF
                    is_white = font_color in ('00FFFFFF', 'FFFFFFFF')
                    if is_bold and is_white:
                        text_pass += 1
                except Exception:
                    pass

        if text_pass == text_total:
            print(f"PASS: Component 3 — All {text_total} milestone cells have bold white text (0.4 pts)")
            total_score += 0.4
        elif text_pass > 0:
            partial = round(0.4 * (text_pass / text_total), 2)
            print(f"PARTIAL: Component 3 — {text_pass}/{text_total} milestone cells have bold white text ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No milestone cells have bold white text")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification (best-effort save for LibreOffice)
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    # Try alternate names
    alt = f'{WORKDIR}/project_plan.xlsx'
    if os.path.exists(alt):
        file_path = alt
    else:
        print(f"File not found: {file_path}")
        print("REWARD: 0.0")
        exit()

persist_app_state()
verify_task(file_path)
