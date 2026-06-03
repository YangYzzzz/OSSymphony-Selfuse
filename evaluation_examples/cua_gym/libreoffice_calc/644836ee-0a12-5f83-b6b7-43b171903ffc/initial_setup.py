"""
Initial Setup: Format weekly project tracker with merged title and milestone rows
Task ID: calc_gsd_020
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timeline"

    # Row 1: Merged title (NOT 16pt bold - task asks agent to do that)
    ws.merge_cells("A1:G1")
    ws["A1"] = "Project Athena - Implementation Plan"
    ws["A1"].font = Font(size=11)  # default size, not 16pt bold
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Row 2: Headers
    headers = ["Task ID", "Task Name", "Owner", "Start", "End", "Duration", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)
        ws.cell(row=2, column=col).font = Font(bold=True)

    # Project task data (rows 3-50)
    # Milestone rows at absolute rows 7, 15, 24, 33, 44
    owners = [
        "Sarah Chen", "Marcus Johnson", "Priya Patel", "James Wilson",
        "Emily Zhang", "Carlos Rivera", "Anna Kowalski", "David Kim",
        "Rachel Foster", "Omar Hassan", "Lisa Nakamura", "Thomas Berg",
        "Nina Petrova", "Alex Dubois", "Maria Santos"
    ]
    statuses = ["Not Started", "In Progress", "Completed", "On Hold", "At Risk"]

    milestone_rows = {7, 15, 24, 33, 44}

    # Define milestone entries
    milestones = {
        7:  ("MILESTONE", "Phase 1 Complete - Requirements Finalized", "2025-03-01", "2025-03-01", "0 days"),
        15: ("MILESTONE", "Phase 2 Complete - Design Approved", "2025-04-15", "2025-04-15", "0 days"),
        24: ("MILESTONE", "Phase 3 Complete - Development Done", "2025-06-30", "2025-06-30", "0 days"),
        33: ("MILESTONE", "Phase 4 Complete - Testing Signed Off", "2025-08-15", "2025-08-15", "0 days"),
        44: ("MILESTONE", "Phase 5 Complete - Go-Live Ready", "2025-10-01", "2025-10-01", "0 days"),
    }

    # Regular task definitions
    tasks = [
        # Phase 1: Requirements (rows 3-6)
        ("T-001", "Gather stakeholder requirements", "Sarah Chen", "2025-01-06", "2025-01-17", "10 days", "Completed"),
        ("T-002", "Document business processes", "Marcus Johnson", "2025-01-13", "2025-01-31", "15 days", "Completed"),
        ("T-003", "Define success criteria", "Priya Patel", "2025-02-03", "2025-02-14", "10 days", "Completed"),
        ("T-004", "Review regulatory compliance", "James Wilson", "2025-02-10", "2025-02-28", "15 days", "Completed"),
        # Row 7 = MILESTONE
        # Phase 2: Design (rows 8-14)
        ("T-005", "Create system architecture", "Emily Zhang", "2025-03-03", "2025-03-14", "10 days", "Completed"),
        ("T-006", "Design database schema", "Carlos Rivera", "2025-03-10", "2025-03-21", "10 days", "Completed"),
        ("T-007", "UI/UX wireframes", "Anna Kowalski", "2025-03-10", "2025-03-28", "15 days", "Completed"),
        ("T-008", "API specification", "David Kim", "2025-03-17", "2025-03-28", "10 days", "In Progress"),
        ("T-009", "Security assessment", "Rachel Foster", "2025-03-24", "2025-04-04", "10 days", "In Progress"),
        ("T-010", "Integration planning", "Omar Hassan", "2025-03-31", "2025-04-11", "10 days", "Not Started"),
        ("T-011", "Design review meeting", "Sarah Chen", "2025-04-14", "2025-04-14", "1 day", "Not Started"),
        # Row 15 = MILESTONE
        # Phase 3: Development (rows 16-23)
        ("T-012", "Backend core modules", "Carlos Rivera", "2025-04-21", "2025-05-16", "20 days", "Not Started"),
        ("T-013", "Frontend dashboard", "Anna Kowalski", "2025-04-21", "2025-05-09", "15 days", "Not Started"),
        ("T-014", "Payment processing module", "David Kim", "2025-05-05", "2025-05-23", "15 days", "Not Started"),
        ("T-015", "Notification service", "Emily Zhang", "2025-05-12", "2025-05-30", "15 days", "Not Started"),
        ("T-016", "Report generation engine", "Marcus Johnson", "2025-05-19", "2025-06-06", "15 days", "Not Started"),
        ("T-017", "Data migration scripts", "Omar Hassan", "2025-06-02", "2025-06-13", "10 days", "Not Started"),
        ("T-018", "Third-party integrations", "Rachel Foster", "2025-06-09", "2025-06-27", "15 days", "Not Started"),
        ("T-019", "Code review cycle", "Priya Patel", "2025-06-23", "2025-06-27", "5 days", "Not Started"),
        # Row 24 = MILESTONE
        # Phase 4: Testing (rows 25-32)
        ("T-020", "Unit test suite", "Lisa Nakamura", "2025-07-01", "2025-07-11", "10 days", "Not Started"),
        ("T-021", "Integration testing", "Thomas Berg", "2025-07-07", "2025-07-18", "10 days", "Not Started"),
        ("T-022", "Performance benchmarking", "Nina Petrova", "2025-07-14", "2025-07-25", "10 days", "Not Started"),
        ("T-023", "Security penetration testing", "Rachel Foster", "2025-07-21", "2025-08-01", "10 days", "Not Started"),
        ("T-024", "User acceptance testing", "Alex Dubois", "2025-07-28", "2025-08-08", "10 days", "Not Started"),
        ("T-025", "Bug fix sprint", "Carlos Rivera", "2025-08-04", "2025-08-11", "6 days", "Not Started"),
        ("T-026", "Regression testing", "Lisa Nakamura", "2025-08-07", "2025-08-13", "5 days", "Not Started"),
        ("T-027", "QA sign-off", "Maria Santos", "2025-08-14", "2025-08-14", "1 day", "Not Started"),
        # Row 33 = MILESTONE
        # Phase 5: Deployment (rows 34-43)
        ("T-028", "Production environment setup", "Omar Hassan", "2025-08-18", "2025-08-29", "10 days", "Not Started"),
        ("T-029", "Data migration dry run", "David Kim", "2025-08-25", "2025-09-05", "10 days", "Not Started"),
        ("T-030", "Staff training sessions", "Alex Dubois", "2025-09-01", "2025-09-12", "10 days", "Not Started"),
        ("T-031", "Documentation finalization", "Priya Patel", "2025-09-01", "2025-09-12", "10 days", "Not Started"),
        ("T-032", "Rollback plan preparation", "Emily Zhang", "2025-09-08", "2025-09-12", "5 days", "Not Started"),
        ("T-033", "Pilot launch (10% traffic)", "Thomas Berg", "2025-09-15", "2025-09-19", "5 days", "Not Started"),
        ("T-034", "Monitor pilot metrics", "Nina Petrova", "2025-09-15", "2025-09-26", "10 days", "Not Started"),
        ("T-035", "Full production deployment", "Carlos Rivera", "2025-09-22", "2025-09-26", "5 days", "Not Started"),
        ("T-036", "Post-launch monitoring", "Marcus Johnson", "2025-09-29", "2025-09-30", "2 days", "Not Started"),
        ("T-037", "Stakeholder demo and handoff", "Sarah Chen", "2025-09-30", "2025-09-30", "1 day", "Not Started"),
        # Row 44 = MILESTONE
        # Post-launch (rows 45-50)
        ("T-038", "30-day performance review", "Maria Santos", "2025-10-06", "2025-10-17", "10 days", "Not Started"),
        ("T-039", "Customer feedback analysis", "Anna Kowalski", "2025-10-13", "2025-10-24", "10 days", "Not Started"),
        ("T-040", "Optimization sprint 1", "David Kim", "2025-10-20", "2025-11-07", "15 days", "Not Started"),
        ("T-041", "Knowledge base updates", "Priya Patel", "2025-10-27", "2025-11-07", "10 days", "Not Started"),
        ("T-042", "Phase 6 planning kickoff", "Sarah Chen", "2025-11-03", "2025-11-14", "10 days", "Not Started"),
        ("T-043", "Archive project artifacts", "James Wilson", "2025-11-10", "2025-11-14", "5 days", "Not Started"),
    ]

    # Write data rows
    data_row = 3
    task_idx = 0
    for row_num in range(3, 51):
        if row_num in milestone_rows:
            ms = milestones[row_num]
            ws.cell(row=row_num, column=1, value=ms[0])  # "MILESTONE"
            ws.cell(row=row_num, column=2, value=ms[1])  # milestone name
            ws.cell(row=row_num, column=3, value="Program Office")  # owner
            ws.cell(row=row_num, column=4, value=ms[2])  # start
            ws.cell(row=row_num, column=5, value=ms[3])  # end
            ws.cell(row=row_num, column=6, value=ms[4])  # duration
            ws.cell(row=row_num, column=7, value="Pending")  # status
        else:
            if task_idx < len(tasks):
                t = tasks[task_idx]
                for c, val in enumerate(t, 1):
                    ws.cell(row=row_num, column=c, value=val)
                task_idx += 1

    # Set column widths for readability
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
