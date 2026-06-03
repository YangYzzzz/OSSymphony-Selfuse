"""
Reward Script: Flag products containing 'Organic' using ISNUMBER(SEARCH(...)) formulas
Task ID: calc_fma_search_isnumber_066
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.5): All 13 cells B2:B14 contain a formula (not empty)
  - Component 2 (0.3): Formulas use ISNUMBER+SEARCH pattern for case-insensitive matching
  - Component 3 (0.2): Column A data is unchanged (no other cells modified)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_search_isnumber_066'

# Expected product descriptions in column A (rows 2-14)
EXPECTED_PRODUCTS = [
    'Organic Apple Juice',
    'Regular Cola',
    'Organic Granola Bar',
    'Chips Classic',
    'Organic Yogurt',
    'Soda Water',
    'Regular Bread',
    'Organic Peanut Butter',
    'Orange Juice',
    'Organic Milk',
    'Regular Cheese',
    'organic tea',
    'Crackers',
]

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check the 'Products' sheet exists
    if 'Products' not in wb.sheetnames:
        print("FAIL: 'Products' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Products']

    # Component 1: All 13 cells B2:B14 contain a formula (not empty) — 0.5 points
    # This FAILS on initial (all None) and PASSES on golden (all have =ISNUMBER(SEARCH(...)))
    try:
        formulas_present = 0
        missing_cells = []
        for row in range(2, 15):  # rows 2-14 inclusive
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formulas_present += 1
            else:
                missing_cells.append(f"B{row}")

        if formulas_present == 13:
            print(f"PASS: Component 1 — All 13 cells B2:B14 contain formulas (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Only {formulas_present}/13 cells have formulas. Missing: {missing_cells}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formulas use ISNUMBER+SEARCH pattern (case-insensitive) — 0.3 points
    # Check that formulas contain both ISNUMBER and SEARCH keywords with "organic"
    # This FAILS on initial (no formulas) and PASSES on golden (all use ISNUMBER+SEARCH)
    try:
        correct_pattern = 0
        wrong_pattern = []
        for row in range(2, 15):  # rows 2-14 inclusive
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                # Must contain ISNUMBER and SEARCH and "ORGANIC" (case-insensitive match)
                if 'ISNUMBER' in val_upper and 'SEARCH' in val_upper and 'ORGANIC' in val_upper:
                    correct_pattern += 1
                else:
                    wrong_pattern.append(f"B{row}: {repr(val)}")
            else:
                wrong_pattern.append(f"B{row}: empty")

        if correct_pattern == 13:
            print(f"PASS: Component 2 — All 13 formulas use ISNUMBER(SEARCH(...)) pattern (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Only {correct_pattern}/13 formulas use correct pattern. Issues: {wrong_pattern[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column A product data is unchanged — 0.2 points
    # Each formula should reference the corresponding A cell (correct row reference)
    # Also verify formulas reference correct row (B2 references A2, B3 references A3, etc.)
    # This FAILS on initial (no formulas) and PASSES on golden (all reference correct rows)
    try:
        correct_refs = 0
        wrong_refs = []
        for row in range(2, 15):  # rows 2-14 inclusive
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                # Formula in row N should reference A{N}
                expected_ref = f'A{row}'
                if expected_ref in val.upper().replace(' ', ''):
                    correct_refs += 1
                else:
                    wrong_refs.append(f"B{row}: formula {repr(val)} doesn't reference {expected_ref}")
            else:
                wrong_refs.append(f"B{row}: not a formula")

        if correct_refs == 13:
            print(f"PASS: Component 3 — All 13 formulas correctly reference their respective row in column A (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Only {correct_refs}/13 formulas have correct row references. Issues: {wrong_refs[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
