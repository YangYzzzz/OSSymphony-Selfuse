"""
Initial Setup: Products spreadsheet with product descriptions in column A and empty 'Is Organic?' column B
Task ID: calc_fma_search_isnumber_066
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_search_isnumber_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Products ---
    ws = wb.active
    ws.title = 'Products'

    # Headers
    ws['A1'] = 'Product Description'
    ws['B1'] = 'Is Organic?'

    # Product data (rows 2-14) — exactly 13 products as specified in context
    products = [
        'Organic Apple Juice',
        'Regular Cola',
        'Organic Granola Bar',
        'Chips Classic',
        'Organic Yogurt',
        'Soda Water',
        'Regular Bread',
        'Organic Peanut Butter',
        'Orange Juice',
        'Organic Milk',
        'Regular Cheese',
        'organic tea',
        'Crackers',
    ]

    for i, product in enumerate(products, start=2):
        ws.cell(row=i, column=1, value=product)
        # Column B (Is Organic?) left EMPTY — agent must fill with ISNUMBER(SEARCH(...))

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Products')
    print(f'Rows 2-14: 13 product descriptions in column A')
    print(f'Column B (Is Organic?): EMPTY — awaiting formulas')


create_initial()
