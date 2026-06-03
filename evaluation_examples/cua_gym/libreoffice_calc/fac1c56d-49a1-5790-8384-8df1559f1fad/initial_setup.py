"""
Initial Setup: Data entry form validation layout
Task ID: calc_nrv_080
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Form"

    # --- Title row ---
    ws.merge_cells("A1:B1")
    ws["A1"] = "Data Entry Form"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # --- Field labels in column A ---
    labels = {
        2: "Name",
        3: "Age",
        4: "Email",
        5: "Start Date",
        6: "Role",
    }

    label_font = Font(name="Calibri", size=11, bold=True)
    label_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for row, label in labels.items():
        cell_a = ws.cell(row=row, column=1, value=label)
        cell_a.font = label_font
        cell_a.fill = label_fill
        cell_a.alignment = Alignment(vertical="center")
        cell_a.border = thin_border

        # B column cells - empty, with border for input area look
        cell_b = ws.cell(row=row, column=2)
        cell_b.border = thin_border
        cell_b.alignment = Alignment(vertical="center")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 35

    # --- Row heights ---
    ws.row_dimensions[1].height = 30
    for r in range(2, 7):
        ws.row_dimensions[r].height = 25

    # --- Instructions area ---
    ws["A8"] = "Instructions:"
    ws["A8"].font = Font(name="Calibri", size=10, bold=True, italic=True)
    ws["A9"] = "Fill in all fields above. Each field has specific requirements."
    ws["A9"].font = Font(name="Calibri", size=9, color="666666")
    ws["A10"] = "Name: 2-50 characters. Age: 18-120. Email: must contain @."
    ws["A10"].font = Font(name="Calibri", size=9, color="666666")
    ws["A11"] = "Start Date: after Jan 1, 2020. Role: select from dropdown."
    ws["A11"].font = Font(name="Calibri", size=9, color="666666")

    # NO data validations on B2:B6 — that is the task

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
