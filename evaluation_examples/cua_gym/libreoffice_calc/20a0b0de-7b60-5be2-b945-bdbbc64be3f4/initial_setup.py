"""
Initial Setup: Insert Pivot Analysis sheet, reorder sheets, apply tab color
Task ID: calc_ggf_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Dashboard ---
    ws_dash = wb.active
    ws_dash.title = 'Dashboard'

    # Header
    ws_dash.merge_cells('A1:E1')
    ws_dash['A1'] = 'Q1 2025 Performance Dashboard'
    ws_dash['A1'].font = Font(name='Arial', size=14, bold=True)
    ws_dash['A1'].alignment = Alignment(horizontal='center')

    # KPI headers
    kpi_headers = ['Metric', 'Target', 'Actual', 'Variance', 'Status']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    for col, h in enumerate(kpi_headers, 1):
        cell = ws_dash.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = header_fill

    kpi_data = [
        ['Total Revenue', 1250000, 1387450, 137450, 'Above Target'],
        ['New Customers', 500, 478, -22, 'Below Target'],
        ['Customer Retention', 0.92, 0.945, 0.025, 'Above Target'],
        ['Avg Order Value', 85.00, 91.30, 6.30, 'Above Target'],
        ['Support Tickets', 200, 187, -13, 'Above Target'],
        ['Employee Satisfaction', 4.0, 4.2, 0.2, 'Above Target'],
        ['Net Promoter Score', 60, 67, 7, 'Above Target'],
        ['Website Conversion', 0.035, 0.041, 0.006, 'Above Target'],
    ]
    for r, row_data in enumerate(kpi_data, 4):
        for c, val in enumerate(row_data, 1):
            ws_dash.cell(row=r, column=c, value=val)

    ws_dash.column_dimensions['A'].width = 22
    ws_dash.column_dimensions['B'].width = 14
    ws_dash.column_dimensions['C'].width = 14
    ws_dash.column_dimensions['D'].width = 14
    ws_dash.column_dimensions['E'].width = 16

    # --- Sheet 2: Summary ---
    ws_sum = wb.create_sheet('Summary')
    sum_headers = ['Department', 'Q1 Revenue', 'Q1 Expenses', 'Net Profit', 'Headcount', 'Revenue per Head']
    for col, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    dept_data = [
        ['Engineering', 425000, 310000, 115000, 42, 10119.05],
        ['Sales', 387000, 198000, 189000, 28, 13821.43],
        ['Marketing', 215000, 175000, 40000, 18, 11944.44],
        ['Operations', 178000, 142000, 36000, 22, 8090.91],
        ['Customer Success', 112000, 87000, 25000, 15, 7466.67],
        ['Human Resources', 45000, 52000, -7000, 8, 5625.00],
        ['Finance', 25450, 31000, -5550, 6, 4241.67],
    ]
    for r, row_data in enumerate(dept_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_sum.cell(row=r, column=c, value=val)

    ws_sum.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_sum.column_dimensions[col_letter].width = 16

    # --- Sheet 3: Raw ---
    ws_raw = wb.create_sheet('Raw')
    raw_headers = ['Transaction ID', 'Date', 'Customer Name', 'Product', 'Category', 'Quantity', 'Unit Price', 'Total', 'Region', 'Sales Rep']
    for col, h in enumerate(raw_headers, 1):
        cell = ws_raw.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    raw_data = [
        ['TXN-1001', '2025-01-05', 'Sarah Chen', 'Widget Pro X', 'Hardware', 3, 149.99, 449.97, 'West', 'James Miller'],
        ['TXN-1002', '2025-01-07', 'Marcus Johnson', 'CloudSync License', 'Software', 1, 299.00, 299.00, 'East', 'Emily Park'],
        ['TXN-1003', '2025-01-08', 'Priya Sharma', 'DataFlow Hub', 'Software', 2, 189.50, 379.00, 'South', 'Carlos Mendez'],
        ['TXN-1004', '2025-01-10', 'David O\'Brien', 'Widget Pro X', 'Hardware', 5, 149.99, 749.95, 'Midwest', 'James Miller'],
        ['TXN-1005', '2025-01-12', 'Aiko Tanaka', 'Support Package', 'Services', 1, 500.00, 500.00, 'West', 'Emily Park'],
        ['TXN-1006', '2025-01-15', 'Elena Volkov', 'DataFlow Hub', 'Software', 3, 189.50, 568.50, 'East', 'Carlos Mendez'],
        ['TXN-1007', '2025-01-18', 'Robert Kim', 'Widget Basic', 'Hardware', 10, 49.99, 499.90, 'South', 'James Miller'],
        ['TXN-1008', '2025-01-20', 'Fatima Al-Hassan', 'CloudSync License', 'Software', 4, 299.00, 1196.00, 'West', 'Emily Park'],
        ['TXN-1009', '2025-01-22', 'Thomas Mueller', 'Support Package', 'Services', 2, 500.00, 1000.00, 'Midwest', 'Carlos Mendez'],
        ['TXN-1010', '2025-01-25', 'Lisa Nakamura', 'Widget Pro X', 'Hardware', 1, 149.99, 149.99, 'East', 'James Miller'],
        ['TXN-1011', '2025-01-28', 'James Wright', 'DataFlow Hub', 'Software', 1, 189.50, 189.50, 'South', 'Emily Park'],
        ['TXN-1012', '2025-02-01', 'Maria Santos', 'Widget Basic', 'Hardware', 8, 49.99, 399.92, 'West', 'Carlos Mendez'],
        ['TXN-1013', '2025-02-04', 'Oliver Chang', 'CloudSync License', 'Software', 2, 299.00, 598.00, 'Midwest', 'James Miller'],
        ['TXN-1014', '2025-02-07', 'Anna Kowalski', 'Support Package', 'Services', 1, 500.00, 500.00, 'East', 'Emily Park'],
        ['TXN-1015', '2025-02-10', 'Raj Patel', 'Widget Pro X', 'Hardware', 4, 149.99, 599.96, 'South', 'Carlos Mendez'],
    ]
    for r, row_data in enumerate(raw_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_raw.cell(row=r, column=c, value=val)

    ws_raw.column_dimensions['A'].width = 14
    ws_raw.column_dimensions['B'].width = 12
    ws_raw.column_dimensions['C'].width = 20
    ws_raw.column_dimensions['D'].width = 18
    ws_raw.column_dimensions['E'].width = 12
    ws_raw.column_dimensions['F'].width = 10
    ws_raw.column_dimensions['G'].width = 12
    ws_raw.column_dimensions['H'].width = 12
    ws_raw.column_dimensions['I'].width = 10
    ws_raw.column_dimensions['J'].width = 16

    ws_raw.auto_filter.ref = 'A1:J16'

    # --- Sheet 4: Charts ---
    ws_charts = wb.create_sheet('Charts')
    chart_headers = ['Month', 'Revenue', 'Expenses', 'Profit']
    for col, h in enumerate(chart_headers, 1):
        cell = ws_charts.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    chart_data = [
        ['January', 425000, 310000, 115000],
        ['February', 462000, 328000, 134000],
        ['March', 500450, 345000, 155450],
    ]
    for r, row_data in enumerate(chart_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_charts.cell(row=r, column=c, value=val)

    ws_charts.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D']:
        ws_charts.column_dimensions[col_letter].width = 14

    # Verify sheet order before save
    assert wb.sheetnames == ['Dashboard', 'Summary', 'Raw', 'Charts'], \
        f"Sheet order mismatch: {wb.sheetnames}"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
