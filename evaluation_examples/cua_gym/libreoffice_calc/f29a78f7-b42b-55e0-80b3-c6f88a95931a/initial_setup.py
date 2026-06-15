"""
Initial Setup: Z-score normalization task for ML features spreadsheet
Task ID: calc_gg5_036
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Features'

    # --- Row 1: Headers ---
    headers = [
        'ObsID', 'Label',
        'Height', 'Weight', 'Age', 'BloodPressure',
        'Cholesterol', 'HeartRate', 'GlucoseLevel', 'BMI'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Feature labels for classification context ---
    labels = ['Healthy', 'At Risk', 'Moderate', 'Critical']

    # --- Rows 2-201: 200 observations with realistic data ---
    for r in range(2, 202):
        obs_id = f'OBS-{r - 1:04d}'
        ws.cell(row=r, column=1, value=obs_id)
        ws.cell(row=r, column=2, value=random.choice(labels))

        # C: Height (cm) ~ 150-195
        height = round(random.gauss(170, 10), 1)
        height = max(145.0, min(200.0, height))
        ws.cell(row=r, column=3, value=height)

        # D: Weight (kg) ~ 50-120
        weight = round(random.gauss(75, 15), 1)
        weight = max(45.0, min(130.0, weight))
        ws.cell(row=r, column=4, value=weight)

        # E: Age ~ 18-80
        age = int(random.gauss(45, 14))
        age = max(18, min(85, age))
        ws.cell(row=r, column=5, value=age)

        # F: BloodPressure (systolic) ~ 90-180
        bp = int(random.gauss(125, 18))
        bp = max(85, min(190, bp))
        ws.cell(row=r, column=6, value=bp)

        # G: Cholesterol (mg/dL) ~ 120-300
        chol = round(random.gauss(200, 40), 1)
        chol = max(110.0, min(320.0, chol))
        ws.cell(row=r, column=7, value=chol)

        # H: HeartRate (bpm) ~ 55-110
        hr = int(random.gauss(78, 12))
        hr = max(50, min(115, hr))
        ws.cell(row=r, column=8, value=hr)

        # I: GlucoseLevel (mg/dL) ~ 60-200
        glucose = round(random.gauss(110, 30), 1)
        glucose = max(55.0, min(220.0, glucose))
        ws.cell(row=r, column=9, value=glucose)

        # J: BMI ~ 16-40
        bmi = round(random.gauss(25, 5), 2)
        bmi = max(15.0, min(42.0, bmi))
        ws.cell(row=r, column=10, value=bmi)

    # Widen columns for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
