"""
Reward Script: Z-score normalization in ml_features.xlsx
Task ID: calc_gg5_036
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Mean row (202) with AVERAGE formulas in C-J
  Component 2 (0.25): StdDev row (203) with STDEV formulas in C-J
  Component 3 (0.15): Z-score column headers in L1:S1 (z_FeatureName)
  Component 4 (0.35): Z-score formulas in L2:S201 using mixed references
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_036'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Features' sheet must exist
    if 'Features' not in wb.sheetnames:
        print("CRITICAL: 'Features' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Features']

    # Map of source columns (C=3 through J=10) to expected feature names
    # Read original headers from row 1 for columns C-J
    source_cols = list(range(3, 11))  # columns C(3) through J(10)
    original_headers = {}
    for c in source_cols:
        h = ws.cell(row=1, column=c).value
        if h is not None:
            original_headers[c] = str(h)

    # ------------------------------------------------------------------
    # Component 1: Mean row (row 202) with AVERAGE formulas (0.25 points)
    # ------------------------------------------------------------------
    try:
        mean_label = ws.cell(row=202, column=1).value
        label_ok = mean_label is not None and str(mean_label).strip().lower() == 'mean'

        avg_count = 0
        for c in source_cols:
            val = ws.cell(row=202, column=c).value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(' ', '')
                # Accept AVERAGE formulas referencing the data range (C2:C201 or similar)
                if 'AVERAGE(' in normalized:
                    avg_count += 1

        if label_ok and avg_count == 8:
            print(f"PASS: Component 1 — Mean row: label='{mean_label}', {avg_count}/8 AVERAGE formulas (0.25 pts)")
            total_score += 0.25
        elif label_ok and avg_count >= 4:
            partial = 0.25 * (avg_count / 8)
            print(f"PARTIAL: Component 1 — Mean row: label='{mean_label}', {avg_count}/8 AVERAGE formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Mean row: label='{mean_label}' (expected 'Mean'), AVERAGE formulas={avg_count}/8")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: StdDev row (row 203) with STDEV formulas (0.25 points)
    # ------------------------------------------------------------------
    try:
        stddev_label = ws.cell(row=203, column=1).value
        label_ok = stddev_label is not None and str(stddev_label).strip().lower() in ('stddev', 'stdev', 'std dev', 'standard deviation')

        stdev_count = 0
        for c in source_cols:
            val = ws.cell(row=203, column=c).value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(' ', '')
                # Accept STDEV or STDEV.S formulas
                if 'STDEV(' in normalized or 'STDEV.S(' in normalized:
                    stdev_count += 1

        if label_ok and stdev_count == 8:
            print(f"PASS: Component 2 — StdDev row: label='{stddev_label}', {stdev_count}/8 STDEV formulas (0.25 pts)")
            total_score += 0.25
        elif label_ok and stdev_count >= 4:
            partial = 0.25 * (stdev_count / 8)
            print(f"PARTIAL: Component 2 — StdDev row: label='{stddev_label}', {stdev_count}/8 STDEV formulas ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — StdDev row: label='{stddev_label}' (expected 'StdDev'), STDEV formulas={stdev_count}/8")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Z-score column headers in L1:S1 (0.15 points)
    # Expected: z_Height, z_Weight, z_Age, z_BloodPressure, z_Cholesterol,
    #           z_HeartRate, z_GlucoseLevel, z_BMI
    # ------------------------------------------------------------------
    try:
        # Normalized columns start at L (col 12) through S (col 19)
        norm_cols = list(range(12, 20))
        expected_headers = {}
        for i, c in enumerate(source_cols):
            if c in original_headers:
                expected_headers[norm_cols[i]] = f"z_{original_headers[c]}"

        header_matches = 0
        for c in norm_cols:
            actual = ws.cell(row=1, column=c).value
            if actual is not None and c in expected_headers:
                # Flexible matching: compare lowercase stripped
                if str(actual).strip().lower() == expected_headers[c].lower():
                    header_matches += 1

        if header_matches == 8:
            print(f"PASS: Component 3 — All 8 z-score headers present in L1:S1 (0.15 pts)")
            total_score += 0.15
        elif header_matches >= 4:
            partial = 0.15 * (header_matches / 8)
            print(f"PARTIAL: Component 3 — {header_matches}/8 z-score headers correct ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {header_matches}/8 z-score headers found in L1:S1")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: Z-score formulas in L2:S201 (0.35 points)
    # Expected pattern: =(Cx-C$202)/C$203 or similar with mixed references
    # Must have formulas in all 1600 cells (8 cols x 200 rows)
    # ------------------------------------------------------------------
    try:
        # Map from norm col index to source col letter
        col_letters = {3: 'C', 4: 'D', 5: 'E', 6: 'F', 7: 'G', 8: 'H', 9: 'I', 10: 'J'}

        formula_count = 0
        correct_pattern_count = 0
        total_cells = 200 * 8  # 1600

        # Check a comprehensive sample: first row, last row, and sample middle rows
        # plus count all formulas
        for norm_idx, norm_col in enumerate(norm_cols):
            src_col = source_cols[norm_idx]
            src_letter = col_letters[src_col]

            for row in range(2, 202):
                val = ws.cell(row=row, column=norm_col).value
                if val is not None and isinstance(val, str) and val.startswith('='):
                    formula_count += 1
                    # Check that the formula references the source column and stats rows
                    upper_val = val.upper().replace(' ', '')
                    # Pattern: references to the source column letter with row anchors ($)
                    # Accept various forms: =(C2-C$202)/C$203 or =(C2-$C$202)/$C$203 etc.
                    if src_letter.upper() in upper_val and '$' in upper_val:
                        # Verify it looks like a z-score: (data - mean) / stdev
                        if '/' in upper_val and ('-' in upper_val or '\u2212' in upper_val):
                            correct_pattern_count += 1

        if correct_pattern_count == total_cells:
            print(f"PASS: Component 4 — All {total_cells} Z-score formulas correct (0.35 pts)")
            total_score += 0.35
        elif correct_pattern_count >= total_cells * 0.9:
            # Nearly all correct
            ratio = correct_pattern_count / total_cells
            partial = 0.35 * ratio
            print(f"PARTIAL: Component 4 — {correct_pattern_count}/{total_cells} Z-score formulas correct ({partial:.3f} pts)")
            total_score += partial
        elif formula_count >= total_cells * 0.5:
            # At least half the cells have formulas
            ratio = formula_count / total_cells
            partial = 0.35 * ratio * 0.5  # penalize for wrong pattern
            print(f"PARTIAL: Component 4 — {formula_count} formulas found, {correct_pattern_count} match Z-score pattern ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {formula_count}/{total_cells} formula cells, {correct_pattern_count} match Z-score pattern")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
