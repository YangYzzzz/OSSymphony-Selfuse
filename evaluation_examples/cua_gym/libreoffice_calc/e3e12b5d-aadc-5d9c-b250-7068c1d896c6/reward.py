"""
Reward Script: Create a pivot in Sheet2 counting customer complaints by product category and resolution outcome
Task ID: osworld_calc_pivot_count_invoice_008
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.3): Summary sheet has a populated pivot table (header row + data rows present)
  - Component 2 (0.4): Correct COUNT values for each product category and resolution outcome
  - Component 3 (0.3): Correct percentage-of-row columns present and accurate
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_count_invoice_008'

# Expected ground truth from the golden file (derived from task description and VM exploration)
# 5 product categories x 4 resolution outcomes
EXPECTED_COUNTS = {
    'Clothing':       {'Resolved': 2, 'Escalated': 2, 'Pending': 1, 'Refunded': 2},
    'Electronics':    {'Resolved': 3, 'Escalated': 2, 'Pending': 2, 'Refunded': 2},
    'Food & Beverage':{'Resolved': 2, 'Escalated': 2, 'Pending': 1, 'Refunded': 1},
    'Home Appliances':{'Resolved': 2, 'Escalated': 2, 'Pending': 2, 'Refunded': 1},
    'Toys & Games':   {'Resolved': 2, 'Escalated': 1, 'Pending': 1, 'Refunded': 2},
}

# Expected percentage of row values (count / row_total), rounded to 4 decimal places
EXPECTED_PCTS = {
    'Clothing':        {'Resolved': 0.2857, 'Escalated': 0.2857, 'Pending': 0.1429, 'Refunded': 0.2857},
    'Electronics':     {'Resolved': 0.3333, 'Escalated': 0.2222, 'Pending': 0.2222, 'Refunded': 0.2222},
    'Food & Beverage': {'Resolved': 0.3333, 'Escalated': 0.3333, 'Pending': 0.1667, 'Refunded': 0.1667},
    'Home Appliances': {'Resolved': 0.2857, 'Escalated': 0.2857, 'Pending': 0.2857, 'Refunded': 0.1429},
    'Toys & Games':    {'Resolved': 0.3333, 'Escalated': 0.1667, 'Pending': 0.1667, 'Refunded': 0.3333},
}

PCT_TOLERANCE = 0.01  # allow 1% tolerance for rounding differences

def parse_pivot_table(ws):
    """
    Parse the pivot table from the Summary sheet.
    Returns a dict: {category: {outcome: count}} and {category: {outcome: pct}}
    Also returns the header row columns mapping.
    """
    counts = {}
    pcts = {}

    # Find the header row: look for a row that has EXACTLY 'Product Category' as the first cell value
    # (not a title row that mentions Product Category within a longer string)
    header_row_idx = None
    for row_idx in range(1, ws.max_row + 1):
        first_cell = ws.cell(row=row_idx, column=1).value
        if first_cell and str(first_cell).strip() == 'Product Category':
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        return None, None, None

    # Parse the header to get column positions
    headers = []
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row_idx, column=col_idx).value
        headers.append(val)

    # Find columns for each outcome (COUNT and %)
    # Headers expected: 'COUNT Resolved', '% of Row (Resolved)', etc.
    count_cols = {}  # outcome -> column index
    pct_cols = {}    # outcome -> column index
    for col_idx, header in enumerate(headers, 1):
        if not header:
            continue
        h = str(header).strip()
        for outcome in ['Resolved', 'Escalated', 'Pending', 'Refunded']:
            # Match COUNT columns: 'COUNT Resolved' or 'Count Resolved' etc.
            if outcome in h and ('COUNT' in h.upper() or 'count' in h.lower()):
                if '%' not in h and 'percent' not in h.lower():
                    count_cols[outcome] = col_idx
            # Match percentage columns: '% of Row (Resolved)' etc.
            if outcome in h and ('%' in h or 'percent' in h.lower() or 'pct' in h.lower()):
                pct_cols[outcome] = col_idx

    # Parse data rows (skip Grand Total row)
    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        category = ws.cell(row=row_idx, column=1).value
        if not category:
            continue
        cat_str = str(category).strip()
        if 'grand' in cat_str.lower() or 'total' in cat_str.lower():
            continue

        counts[cat_str] = {}
        pcts[cat_str] = {}

        for outcome, col_idx in count_cols.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                try:
                    counts[cat_str][outcome] = int(float(val))
                except (ValueError, TypeError):
                    counts[cat_str][outcome] = None

        for outcome, col_idx in pct_cols.items():
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                try:
                    pcts[cat_str][outcome] = float(val)
                except (ValueError, TypeError):
                    pcts[cat_str][outcome] = None

    return counts, pcts, header_row_idx


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary/Sheet2 exists
    summary_sheet = None
    for name in wb.sheetnames:
        if name.lower() in ('summary', 'sheet2'):
            summary_sheet = wb[name]
            break

    if summary_sheet is None:
        print(f"FAIL: No Summary/Sheet2 found. Sheets present: {wb.sheetnames}")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Summary sheet has a populated pivot table header and data rows (0.3 points)
    # This FAILS on initial (empty sheet) -> PASSES on golden (populated pivot table)
    try:
        counts, pcts, header_row_idx = parse_pivot_table(summary_sheet)

        if header_row_idx is not None and counts and len(counts) >= 1:
            print(f"PASS: Component 1 — Summary sheet has pivot data, header at row {header_row_idx}, "
                  f"{len(counts)} product category rows found (0.3 pts)")
            total_score += 0.3
        else:
            if header_row_idx is None:
                print("FAIL: Component 1 — No header row with 'Product Category' found in Summary sheet")
            elif not counts:
                print("FAIL: Component 1 — Summary sheet has header but no data rows")
            else:
                print(f"FAIL: Component 1 — Only {len(counts)} category rows found, expected >= 1")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        counts, pcts, header_row_idx = None, None, None

    # Component 2: Correct COUNT values for all 5 product categories x 4 resolution outcomes (0.4 points)
    # This FAILS on initial (no data) -> PASSES on golden (accurate count values)
    try:
        if counts is None:
            print("FAIL: Component 2 — Cannot check counts, pivot table parse failed")
        else:
            all_categories_correct = True
            correct_cats = 0
            total_expected_cats = len(EXPECTED_COUNTS)

            for cat, expected_outcomes in EXPECTED_COUNTS.items():
                if cat not in counts:
                    print(f"FAIL: Component 2 — Category '{cat}' not found in pivot. Found: {list(counts.keys())}")
                    all_categories_correct = False
                    continue

                cat_correct = True
                for outcome, expected_count in expected_outcomes.items():
                    actual_count = counts[cat].get(outcome)
                    if actual_count != expected_count:
                        print(f"FAIL: Component 2 — {cat}/{outcome}: expected {expected_count}, got {actual_count}")
                        cat_correct = False
                        all_categories_correct = False

                if cat_correct:
                    correct_cats += 1

            if all_categories_correct:
                print(f"PASS: Component 2 — All 5 categories with correct COUNT values for all 4 outcomes (0.4 pts)")
                total_score += 0.4
            elif correct_cats > 0:
                # Partial: each correctly verified category is worth a proportional fraction
                partial = round((correct_cats / total_expected_cats) * 0.4, 4)
                print(f"PARTIAL: Component 2 — {correct_cats}/{total_expected_cats} categories with correct counts "
                      f"({partial:.4f} pts)")
                total_score += partial
            else:
                print("FAIL: Component 2 — No categories with correct count values found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Percentage-of-row columns present and values correct (0.3 points)
    # This FAILS on initial (no data) -> PASSES on golden (percentage columns with correct values)
    try:
        if pcts is None:
            print("FAIL: Component 3 — Cannot check percentages, pivot table parse failed")
        else:
            # Check that percentage columns exist at all
            has_pct_columns = any(len(p) > 0 for p in pcts.values()) if pcts else False

            if not has_pct_columns:
                print("FAIL: Component 3 — No percentage columns found in Summary sheet")
            else:
                all_pcts_correct = True
                correct_pct_cats = 0
                total_expected_cats = len(EXPECTED_PCTS)

                for cat, expected_pct_outcomes in EXPECTED_PCTS.items():
                    if cat not in pcts:
                        print(f"FAIL: Component 3 — Category '{cat}' missing from pct data")
                        all_pcts_correct = False
                        continue

                    cat_correct = True
                    for outcome, expected_pct in expected_pct_outcomes.items():
                        actual_pct = pcts[cat].get(outcome)
                        if actual_pct is None:
                            print(f"FAIL: Component 3 — {cat}/{outcome}: pct value missing")
                            cat_correct = False
                            all_pcts_correct = False
                        elif abs(actual_pct - expected_pct) > PCT_TOLERANCE:
                            print(f"FAIL: Component 3 — {cat}/{outcome}: expected pct ~{expected_pct}, got {actual_pct}")
                            cat_correct = False
                            all_pcts_correct = False

                    if cat_correct:
                        correct_pct_cats += 1

                if all_pcts_correct:
                    print(f"PASS: Component 3 — All percentage columns present and correct for all categories (0.3 pts)")
                    total_score += 0.3
                elif correct_pct_cats > 0:
                    partial = round((correct_pct_cats / total_expected_cats) * 0.3, 4)
                    print(f"PARTIAL: Component 3 — {correct_pct_cats}/{total_expected_cats} categories with correct "
                          f"percentages ({partial:.4f} pts)")
                    total_score += partial
                else:
                    print("FAIL: Component 3 — No categories with correct percentage values found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
