"""
Initial Setup: Create customer complaint records spreadsheet for pivot task
Task ID: osworld_calc_pivot_count_invoice_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_pivot_count_invoice_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Complaints ---
    ws1 = wb.active
    ws1.title = 'Complaints'

    # Headers
    headers = ['Complaint ID', 'Product Category', 'Resolution Outcome', 'Date Filed', 'Customer Region']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Realistic complaint data
    data = [
        ['CMP-0001', 'Electronics',     'Resolved',  '2025-01-05', 'North'],
        ['CMP-0002', 'Clothing',         'Escalated', '2025-01-06', 'South'],
        ['CMP-0003', 'Home Appliances',  'Pending',   '2025-01-07', 'East'],
        ['CMP-0004', 'Electronics',      'Refunded',  '2025-01-08', 'West'],
        ['CMP-0005', 'Food & Beverage',  'Resolved',  '2025-01-09', 'North'],
        ['CMP-0006', 'Clothing',         'Resolved',  '2025-01-10', 'South'],
        ['CMP-0007', 'Electronics',      'Escalated', '2025-01-11', 'East'],
        ['CMP-0008', 'Home Appliances',  'Resolved',  '2025-01-12', 'West'],
        ['CMP-0009', 'Food & Beverage',  'Pending',   '2025-01-13', 'North'],
        ['CMP-0010', 'Toys & Games',     'Refunded',  '2025-01-14', 'South'],
        ['CMP-0011', 'Electronics',      'Pending',   '2025-01-15', 'East'],
        ['CMP-0012', 'Clothing',         'Refunded',  '2025-01-16', 'West'],
        ['CMP-0013', 'Home Appliances',  'Escalated', '2025-01-17', 'North'],
        ['CMP-0014', 'Toys & Games',     'Resolved',  '2025-01-18', 'South'],
        ['CMP-0015', 'Food & Beverage',  'Escalated', '2025-01-19', 'East'],
        ['CMP-0016', 'Electronics',      'Resolved',  '2025-01-20', 'West'],
        ['CMP-0017', 'Clothing',         'Pending',   '2025-01-21', 'North'],
        ['CMP-0018', 'Home Appliances',  'Refunded',  '2025-01-22', 'South'],
        ['CMP-0019', 'Toys & Games',     'Escalated', '2025-01-23', 'East'],
        ['CMP-0020', 'Food & Beverage',  'Refunded',  '2025-01-24', 'West'],
        ['CMP-0021', 'Electronics',      'Resolved',  '2025-01-25', 'North'],
        ['CMP-0022', 'Clothing',         'Escalated', '2025-01-26', 'South'],
        ['CMP-0023', 'Home Appliances',  'Pending',   '2025-01-27', 'East'],
        ['CMP-0024', 'Toys & Games',     'Pending',   '2025-01-28', 'West'],
        ['CMP-0025', 'Electronics',      'Refunded',  '2025-01-29', 'North'],
        ['CMP-0026', 'Clothing',         'Resolved',  '2025-01-30', 'South'],
        ['CMP-0027', 'Food & Beverage',  'Resolved',  '2025-01-31', 'East'],
        ['CMP-0028', 'Home Appliances',  'Resolved',  '2025-02-01', 'West'],
        ['CMP-0029', 'Toys & Games',     'Refunded',  '2025-02-02', 'North'],
        ['CMP-0030', 'Electronics',      'Pending',   '2025-02-03', 'South'],
        ['CMP-0031', 'Clothing',         'Refunded',  '2025-02-04', 'East'],
        ['CMP-0032', 'Food & Beverage',  'Escalated', '2025-02-05', 'West'],
        ['CMP-0033', 'Home Appliances',  'Escalated', '2025-02-06', 'North'],
        ['CMP-0034', 'Electronics',      'Escalated', '2025-02-07', 'South'],
        ['CMP-0035', 'Toys & Games',     'Resolved',  '2025-02-08', 'East'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 18

    # --- Sheet 2: Summary (empty, where pivot will be created) ---
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_state = 'visible'
    # Sheet2 is intentionally empty — agent will create the pivot here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
