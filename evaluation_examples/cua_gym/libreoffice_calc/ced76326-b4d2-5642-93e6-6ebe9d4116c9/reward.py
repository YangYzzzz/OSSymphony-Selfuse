"""
Reward Script: Set up department budget tracker with variance, status, totals, and conditional formatting
Task ID: calc_edu_budget_dept_016
Domain: libreoffice_calc
Scoring:
  Component 1: Variance formulas in D2:D16 (=Cx-Bx pattern)        — 0.30 pts
  Component 2: Status formulas in E2:E16 (IF(Dx>0,...) pattern)      — 0.25 pts
  Component 3: Totals row SUM formulas in B17, C17, D17              — 0.20 pts
  Component 4: Conditional formatting orange fill when C > B*1.1     — 0.15 pts
  Component 5: Currency number format on B, C, D columns             — 0.10 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_budget_dept_016'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if 'DeptBudget' not in wb.sheetnames:
        print("CRITICAL: Sheet 'DeptBudget' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DeptBudget']

    # Component 1: Variance formulas in D2:D16 (=Cx-Bx pattern) — 0.30 points
    # Each row should have a formula like =C2-B2, =C3-B3, etc.
    # This FAILS on initial (D2:D16 are empty None) and PASSES on golden.
    try:
        variance_formula_count = 0
        total_rows = 15  # rows 2 through 16
        for row in range(2, 17):
            cell_val = ws.cell(row=row, column=4).value  # Column D
            if cell_val is not None and isinstance(cell_val, str):
                # Normalize formula for comparison: remove spaces, uppercase
                normalized = cell_val.strip().upper().replace(' ', '')
                # Expected pattern: =C{row}-B{row}
                expected = f'=C{row}-B{row}'
                if normalized == expected:
                    variance_formula_count += 1
        if variance_formula_count == total_rows:
            print(f"PASS: Component 1 — All {total_rows} variance formulas present in D2:D16 (0.30 pts)")
            total_score += 0.30
        elif variance_formula_count >= 10:
            print(f"PARTIAL: Component 1 — {variance_formula_count}/{total_rows} variance formulas present (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {variance_formula_count}/{total_rows} variance formulas found in D2:D16")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Status formulas in E2:E16 (IF(D>0,...) pattern) — 0.25 points
    # Each row should have a formula like =IF(D2>0,"Over Budget","Under Budget")
    # This FAILS on initial (E2:E16 are empty None) and PASSES on golden.
    try:
        status_formula_count = 0
        total_rows = 15  # rows 2 through 16
        for row in range(2, 17):
            cell_val = ws.cell(row=row, column=5).value  # Column E
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.strip().upper().replace(' ', '')
                # Expected pattern: =IF(D{row}>0,"OVER BUDGET","UNDER BUDGET")
                expected = f'=IF(D{row}>0,"OVERBUDGET","UNDERBUDGET")'
                if normalized == expected:
                    status_formula_count += 1
        if status_formula_count == total_rows:
            print(f"PASS: Component 2 — All {total_rows} status formulas present in E2:E16 (0.25 pts)")
            total_score += 0.25
        elif status_formula_count >= 10:
            print(f"PARTIAL: Component 2 — {status_formula_count}/{total_rows} status formulas present (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — Only {status_formula_count}/{total_rows} status formulas found in E2:E16")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Totals row SUM formulas in B17, C17, D17 — 0.20 points
    # Row 17 should contain SUM formulas for columns B, C, and D.
    # This FAILS on initial (B17 has text 'Total Budget', C17/D17 are None) and PASSES on golden.
    try:
        b17 = ws.cell(row=17, column=2).value  # B17
        c17 = ws.cell(row=17, column=3).value  # C17
        d17 = ws.cell(row=17, column=4).value  # D17

        b17_ok = (isinstance(b17, str) and '=SUM(B2:B16)' in b17.upper().replace(' ', ''))
        c17_ok = (isinstance(c17, str) and '=SUM(C2:C16)' in c17.upper().replace(' ', ''))
        d17_ok = (isinstance(d17, str) and '=SUM(D2:D16)' in d17.upper().replace(' ', ''))

        sums_count = sum([b17_ok, c17_ok, d17_ok])
        if sums_count == 3:
            print(f"PASS: Component 3 — SUM formulas in B17, C17, D17 all present (0.20 pts)")
            total_score += 0.20
        elif sums_count == 2:
            print(f"PARTIAL: Component 3 — {sums_count}/3 SUM formulas present in row 17 (0.10 pts)")
            total_score += 0.10
        elif sums_count == 1:
            print(f"PARTIAL: Component 3 — Only {sums_count}/3 SUM formulas in row 17 (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No SUM formulas found in B17/C17/D17. Got B17={repr(b17)}, C17={repr(c17)}, D17={repr(d17)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting orange fill (#FFA500) when C > B*1.1 — 0.15 points
    # A formula-based CF rule covering A2:E16 (or similar covering data range)
    # with fill color FFFFA500 and formula $C>$B*1.1.
    # This FAILS on initial (no CF rules) and PASSES on golden.
    try:
        cf_found = False
        orange_color_found = False
        correct_formula_found = False

        for cf_range_obj, rules in ws.conditional_formatting._cf_rules.items():
            for rule in rules:
                # Check for fill with orange color
                if hasattr(rule, 'dxf') and rule.dxf and hasattr(rule.dxf, 'fill') and rule.dxf.fill:
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb
                        # Orange can be FFFFA500 exactly
                        if fill_color and 'FFA500' in fill_color.upper():
                            orange_color_found = True
                    except Exception:
                        pass

                # Check for formula containing C > B * 1.1 pattern
                if hasattr(rule, 'formula') and rule.formula:
                    for formula in rule.formula:
                        formula_upper = str(formula).upper().replace(' ', '')
                        # Look for C > B*1.1 pattern (various equivalent forms)
                        if ('C' in formula_upper and 'B' in formula_upper and '1.1' in formula_upper):
                            correct_formula_found = True
                        elif ('C' in formula_upper and 'B' in formula_upper and '110' in formula_upper):
                            # Alternative: C > B*110/100 or similar
                            correct_formula_found = True

                cf_found = True

        if cf_found and orange_color_found and correct_formula_found:
            print(f"PASS: Component 4 — Conditional formatting with orange (#FFA500) and C>B*1.1 formula present (0.15 pts)")
            total_score += 0.15
        elif cf_found and orange_color_found:
            print(f"PARTIAL: Component 4 — CF rule with orange fill found but formula check inconclusive (0.08 pts)")
            total_score += 0.08
        elif cf_found:
            print(f"PARTIAL: Component 4 — CF rule found but orange fill not confirmed (0.04 pts)")
            total_score += 0.04
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Currency number format on columns B, C, D — 0.10 points
    # Columns B, C, D should be formatted as currency (e.g., $#,##0.00).
    # This FAILS on initial (B/C are 'General', D is 'General') and PASSES on golden.
    try:
        currency_formats = {'$#,##0.00', '#,##0.00', '"$"#,##0.00', '$#,##0', '#,##0'}
        # Check a sample of cells in each column: B2, C2, D2
        b_fmt = ws.cell(row=2, column=2).number_format
        c_fmt = ws.cell(row=2, column=3).number_format
        d_fmt = ws.cell(row=2, column=4).number_format

        b_currency = (b_fmt != 'General' and (b_fmt in currency_formats or '$' in b_fmt or '0.00' in b_fmt))
        c_currency = (c_fmt != 'General' and (c_fmt in currency_formats or '$' in c_fmt or '0.00' in c_fmt))
        d_currency = (d_fmt != 'General' and (d_fmt in currency_formats or '$' in d_fmt or '0.00' in d_fmt))

        currency_count = sum([b_currency, c_currency, d_currency])
        if currency_count == 3:
            print(f"PASS: Component 5 — All currency formats set on B ({b_fmt}), C ({c_fmt}), D ({d_fmt}) columns (0.10 pts)")
            total_score += 0.10
        elif currency_count >= 2:
            print(f"PARTIAL: Component 5 — {currency_count}/3 columns have currency format (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Currency format not set. B={repr(b_fmt)}, C={repr(c_fmt)}, D={repr(d_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
