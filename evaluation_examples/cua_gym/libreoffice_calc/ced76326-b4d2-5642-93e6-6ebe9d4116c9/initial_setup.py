"""
Initial Setup: Department Budget Tracker
Task ID: calc_edu_budget_dept_016
Domain: libreoffice_calc

Creates the initial pre-task spreadsheet with DeptBudget sheet containing
15 budget line items with Category, Budgeted, and Actual columns.
Columns D (Variance) and E (Status) are intentionally empty — the agent
will add those formulas. Row 17 totals are also empty.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_budget_dept_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DeptBudget ---
    ws = wb.active
    ws.title = 'DeptBudget'

    # Headers in row 1
    headers = ['Category', 'Budgeted', 'Actual', 'Variance', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Style headers bold
    bold_font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=1, column=col).font = bold_font

    # 15 realistic budget line items (rows 2-16)
    # Mix of over-budget, under-budget, and >10% over-budget items
    # Columns: Category, Budgeted, Actual  (D and E intentionally empty)
    budget_data = [
        # Category,                  Budgeted,   Actual
        ('Faculty Salaries',         320000,     328500),   # slightly over
        ('Adjunct Instructor Fees',   48000,      44200),   # under
        ('Administrative Staff',      85000,      87300),   # slightly over
        ('Graduate Assistantships',   62000,      68800),   # over >10%
        ('Research Supplies',         25000,      22400),   # under
        ('Lab Equipment',             40000,      45200),   # over >10%
        ('Office Supplies',            8500,       7900),   # under
        ('Software Licenses',         15000,      16350),   # slightly over
        ('Travel & Conferences',      18000,      20100),   # over ~11.7%
        ('Library Resources',         12000,      11500),   # under
        ('Student Services',           9500,       9800),   # slightly over
        ('Facilities & Utilities',    35000,      34100),   # under
        ('IT Infrastructure',         22000,      24800),   # over >10%
        ('Professional Development',  11000,      10300),   # under
        ('Miscellaneous Expenses',     6000,       6450),   # slightly over
    ]

    for r, (category, budgeted, actual) in enumerate(budget_data, 2):
        ws.cell(row=r, column=1, value=category)
        ws.cell(row=r, column=2, value=budgeted)
        ws.cell(row=r, column=3, value=actual)
        # Columns D (Variance) and E (Status) intentionally left empty

    # Row 17: Totals row label only — B17, C17, D17, E17 intentionally empty
    ws.cell(row=17, column=1, value='Total Budget')
    ws.cell(row=17, column=1).font = bold_font

    # Format column widths for readability
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: DeptBudget')
    print('  - 15 budget line items (rows 2-16)')
    print('  - Columns A (Category), B (Budgeted), C (Actual) filled')
    print('  - Columns D (Variance) and E (Status) intentionally empty')
    print('  - Row 17: Total Budget label only, no formulas')


create_initial()
