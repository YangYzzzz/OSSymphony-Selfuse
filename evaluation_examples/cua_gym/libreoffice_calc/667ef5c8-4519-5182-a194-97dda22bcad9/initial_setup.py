"""
Initial Setup: Create a Weekly Plan workbook with realistic team scheduling data.
Task ID: calc_ps_036
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Weekly Plan'

    # --- Headers ---
    headers = ['Task', 'Owner', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 22

    # --- Data rows 2-25 (24 rows of realistic weekly plan data) ---
    data = [
        ['Review Q1 marketing metrics', 'Sarah Chen', 'Research data', 'Compile report', 'Present to team', '', ''],
        ['Update product roadmap', 'Marcus Johnson', '', 'Draft updates', 'Review with PM', 'Finalize doc', ''],
        ['Fix login authentication bug', 'Priya Patel', 'Reproduce issue', 'Debug backend', 'Write unit tests', 'Deploy fix', 'Verify in prod'],
        ['Prepare client onboarding deck', 'James Wilson', 'Gather assets', 'Design slides', '', 'Internal review', 'Send to client'],
        ['Conduct user interviews', 'Elena Rodriguez', 'Schedule calls', 'Interview batch 1', 'Interview batch 2', 'Synthesize notes', 'Share findings'],
        ['Optimize database queries', 'David Kim', 'Profile slow queries', 'Rewrite joins', 'Add indexes', 'Load testing', 'Deploy changes'],
        ['Design new landing page', 'Aisha Mohammed', 'Wireframe draft', 'Hi-fi mockup', 'Team feedback', 'Revisions', 'Handoff to dev'],
        ['Write API documentation', 'Tom Bradley', 'Audit endpoints', 'Draft auth docs', 'Draft CRUD docs', 'Peer review', 'Publish to wiki'],
        ['Plan team offsite event', 'Lisa Chang', 'Scout venues', 'Get quotes', 'Book venue', 'Send invites', 'Confirm RSVPs'],
        ['Migrate CI/CD to GitHub Actions', 'Ryan O\'Brien', 'Map current pipeline', 'Write YAML configs', 'Test staging', 'Test production', 'Cutover'],
        ['Implement dark mode toggle', 'Mei Lin', 'Design tokens', 'CSS variables', 'Component updates', 'QA testing', 'Release'],
        ['Audit security permissions', 'Carlos Vega', 'Export user roles', 'Identify gaps', 'Propose changes', 'Implement fixes', 'Final audit'],
        ['Create sales forecast model', 'Jennifer Park', 'Gather historical data', 'Build model', 'Validate accuracy', 'Present to VP', ''],
        ['Refactor payment module', 'Alex Turner', 'Code review', 'Extract services', 'Update interfaces', 'Integration tests', 'Deploy'],
        ['Set up monitoring dashboards', 'Nina Kowalski', 'Define KPIs', 'Grafana setup', 'Alert rules', 'Test alerts', 'Documentation'],
        ['Organize knowledge base', 'Robert Huang', 'Audit old articles', 'Archive outdated', 'Update templates', 'New categories', 'Announce to team'],
        ['Test mobile responsiveness', 'Samantha Lee', 'Test on iOS', 'Test on Android', 'File bug reports', 'Retest fixes', 'Sign-off'],
        ['Prepare quarterly budget', 'Michael Torres', 'Collect dept requests', 'Draft allocations', 'Review with CFO', 'Revise', 'Submit final'],
        ['Integrate Slack notifications', 'Anna Petrova', 'Webhook setup', 'Message templates', 'Channel routing', 'Test end-to-end', 'Go live'],
        ['Run A/B test on checkout flow', 'Chris Nakamura', 'Define variants', 'Implement changes', 'Launch test', 'Monitor metrics', 'Analyze results'],
        ['Update employee handbook', 'Diana Foster', 'Review policies', 'Draft new sections', 'Legal review', 'Format doc', 'Distribute'],
        ['Build customer feedback widget', 'Kevin Singh', 'Design UI component', 'Backend API', 'Frontend integration', 'Accessibility check', 'Deploy'],
        ['Coordinate vendor demo', 'Rachel Goldstein', 'Contact vendors', 'Schedule demos', 'Attend demo 1', 'Attend demo 2', 'Evaluation matrix'],
        ['Performance review prep', 'Brian Walsh', 'Self-assessment', 'Collect peer feedback', 'Draft reviews', 'Calibration meeting', 'Deliver reviews'],
    ]

    data_font = Font(name='Calibri', size=11)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c >= 3:  # Day columns centered
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # No sheet protection - this is the initial state

    # Set row height for better readability
    ws.row_dimensions[1].height = 22
    for r in range(2, 26):
        ws.row_dimensions[r].height = 30

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
