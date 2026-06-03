"""
Reward Script: Calculate asset age with DATEDIF and apply conditional formatting
Task ID: calc_gg5_043
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): DATEDIF formulas in F2:F80
  Component 2 (0.5): Conditional formatting on F2:F80 (orange fill for >5)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_043'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Assets' sheet must exist
    if 'Assets' not in wb.sheetnames:
        print("CRITICAL: 'Assets' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Assets']

    # Component 1: DATEDIF formulas in F2:F80 (0.5 points)
    # Each cell should contain a DATEDIF formula referencing the D column and TODAY()
    try:
        formula_count = 0
        total_cells = 79  # F2 through F80

        for row in range(2, 81):
            cell_val = ws.cell(row=row, column=6).value
            if cell_val and isinstance(cell_val, str):
                # Normalize: strip spaces, uppercase
                normalized = cell_val.upper().replace(" ", "")
                # Check for DATEDIF pattern: =DATEDIF(D<row>,TODAY(),"Y")
                # Allow flexible patterns like =DATEDIF(D2,TODAY(),"Y")
                if "DATEDIF" in normalized and "TODAY()" in normalized and '"Y"' in normalized:
                    # Also verify it references the correct D cell for this row
                    expected_ref = f"D{row}"
                    if expected_ref in normalized:
                        formula_count += 1

        # Award proportional credit based on how many formulas are correct
        formula_ratio = formula_count / total_cells
        if formula_ratio >= 0.95:
            # Nearly all or all formulas correct
            component1_score = 0.5
            print(f"PASS: Component 1 — {formula_count}/{total_cells} DATEDIF formulas correct (0.5 pts)")
        elif formula_ratio >= 0.5:
            component1_score = round(0.5 * formula_ratio, 2)
            print(f"PARTIAL: Component 1 — {formula_count}/{total_cells} DATEDIF formulas ({component1_score} pts)")
        else:
            component1_score = 0.0
            print(f"FAIL: Component 1 — Only {formula_count}/{total_cells} DATEDIF formulas found")

        total_score += component1_score

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Conditional formatting on F2:F80 with orange fill for >5 (0.5 points)
    try:
        cf_rules = ws.conditional_formatting
        found_valid_rule = False

        for cf in cf_rules:
            range_str = str(cf).upper()
            # Check that the range covers F2:F80 (or equivalent)
            covers_f_range = False
            if "F2:F80" in range_str or "F2:F80" in str(cf).replace(" ", ""):
                covers_f_range = True

            if not covers_f_range:
                # Also check individual cell ranges that might cover the F column
                # Some implementations might use $F$2:$F$80
                range_no_dollar = range_str.replace("$", "")
                if "F2:F80" in range_no_dollar:
                    covers_f_range = True

            if covers_f_range:
                for rule in cf.rules:
                    # Check rule type and operator
                    is_greater_than_5 = False
                    if rule.type == 'cellIs' and rule.operator == 'greaterThan':
                        if rule.formula and '5' in str(rule.formula[0]):
                            is_greater_than_5 = True

                    if is_greater_than_5:
                        # Check for orange fill
                        has_orange_fill = False
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            fg = rule.dxf.fill.fgColor
                            if fg and fg.rgb:
                                rgb_str = str(fg.rgb).upper()
                                # Orange colors: FFA500 is pure orange
                                # Accept common orange variants (FFA500, FF8C00, etc.)
                                # Extract the RGB portion (last 6 chars of ARGB)
                                rgb_hex = rgb_str[-6:] if len(rgb_str) >= 6 else rgb_str
                                r_val = int(rgb_hex[0:2], 16)
                                g_val = int(rgb_hex[2:4], 16)
                                b_val = int(rgb_hex[4:6], 16)
                                # Orange: high red, medium green, low blue
                                if r_val >= 200 and g_val >= 100 and g_val <= 200 and b_val <= 50:
                                    has_orange_fill = True

                        if has_orange_fill:
                            found_valid_rule = True
                            print(f"PASS: Component 2 — CF rule found: cellIs greaterThan 5, orange fill on F2:F80 (0.5 pts)")
                            total_score += 0.5
                            break
                        else:
                            # Rule matches >5 but fill is not orange
                            # Give partial credit (0.25) for having the right rule logic
                            print(f"PARTIAL: Component 2 — CF rule >5 found but fill color not orange ({rgb_str if fg and fg.rgb else 'unknown'})")
                            total_score += 0.25
                            found_valid_rule = True
                            break

            if found_valid_rule:
                break

        if not found_valid_rule:
            # Check if any CF rules exist at all on the F column
            any_f_rules = False
            for cf in cf_rules:
                if 'F' in str(cf).upper():
                    any_f_rules = True
                    break
            if any_f_rules:
                print(f"FAIL: Component 2 — CF rules found on F column but none match criteria (greaterThan 5, orange fill, F2:F80)")
            else:
                print(f"FAIL: Component 2 — No conditional formatting found on F column")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
