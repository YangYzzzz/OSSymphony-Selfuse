"""
Initial Setup: Asset registry with 79 assets, Age column empty
Task ID: calc_gg5_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Assets'

    # --- Headers ---
    headers = ['Asset ID', 'Name', 'Category', 'Purchase Date', 'Purchase Cost', 'Age (Years)']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14

    # --- Asset data generation ---
    categories = [
        'IT Equipment', 'Furniture', 'Vehicles', 'Lab Equipment',
        'Office Supplies', 'HVAC Systems', 'Safety Equipment', 'Telecom'
    ]

    asset_names = [
        'Dell OptiPlex 7090 Desktop', 'HP LaserJet Pro M404n', 'Cisco Catalyst 9200 Switch',
        'Ergotron Sit-Stand Desk', 'Herman Miller Aeron Chair', 'Steelcase Leap V2 Chair',
        'Toyota Hilux Pickup 2019', 'Ford Transit Connect Van', 'Chevrolet Express Cargo',
        'Thermo Fisher Centrifuge', 'Agilent HPLC System 1260', 'Shimadzu UV-2600 Spectro',
        'Brother MFC-L8900CDW', 'Canon imageRUNNER C3530i', 'Epson SureColor T5470M',
        'Lenovo ThinkPad X1 Carbon', 'Apple MacBook Pro 16"', 'Microsoft Surface Pro 9',
        'Samsung 27" Curved Monitor', 'LG UltraWide 34WN80C', 'Dell U2722D Monitor',
        'Haworth Zody Task Chair', 'Knoll Generation Chair', 'Global Accord Mesh Chair',
        'Honda CR-V 2020', 'Nissan NV200 Compact Cargo', 'RAM ProMaster City',
        'Beckman Coulter Analyzer', 'PerkinElmer Plate Reader', 'Bio-Rad CFX96 PCR',
        'Xerox VersaLink C405', 'Ricoh MP C3004ex', 'Sharp MX-3071 Copier',
        'Daikin VRV IV Heat Pump', 'Carrier 50XC Chiller Unit', 'Trane Intellipak RTU',
        'Honeywell Fire Panel NFS2', 'Kidde FM-200 System', 'MSA Altair 5X Gas Detector',
        'Avaya IP Office 500 V2', 'Polycom RealPresence Trio', 'Cisco IP Phone 8845',
        'NetApp AFF A250 Storage', 'Dell EMC PowerEdge R750', 'HPE ProLiant DL380 Gen10',
        'APC Smart-UPS SRT 10kVA', 'Eaton 9PX 6000 UPS', 'Tripp Lite SU3000RTXLCD3U',
        'Fluke 87V Multimeter', 'Keysight 34465A DMM', 'Tektronix TBS2000B Scope',
        'Sony FW-75BZ40H Display', 'Samsung QB75R Signage', 'LG 86UH5F Commercial TV',
        'Bosch Security Camera NDE', 'Axis P3245-V Dome Camera', 'Hikvision DS-2CD2347G2',
        'Makita 18V LXT Drill Kit', 'DeWalt 20V MAX Impact Set', 'Milwaukee M18 FUEL Combo',
        'Husqvarna Z254F Zero Turn', 'John Deere Z345M Mower', 'Toro TimeCutter SS4225',
        'Konica Minolta bizhub C258', 'Kyocera TASKalfa 3253ci', 'Lexmark CX860de',
        'Mitsubishi City Multi VRF', 'Lennox XC25 Heat Pump', 'York YC Scroll Chiller',
        'Johnson Controls FX60 BMS', 'Siemens Desigo CC Panel', 'Schneider EcoStruxure BMS',
        'Panasonic KX-NS700 PBX', 'Mitel MiVoice 5330e Phone', 'NEC SL2100 System',
        'Zebra ZT411 Label Printer', 'Datalogic Gryphon Scanner', 'Honeywell Granit 1981i',
        'Caterpillar 320 Excavator', 'Bobcat S590 Skid-Steer',
    ]

    random.seed(43)

    data = []
    for i in range(79):
        asset_id = f'AST-{1001 + i:04d}'
        name = asset_names[i % len(asset_names)]
        category = categories[i % len(categories)]
        # Purchase dates spread from 2015 to 2025
        days_ago = random.randint(180, 4000)
        purchase_date = date.today() - timedelta(days=days_ago)
        purchase_cost = round(random.uniform(250, 185000), 2)
        data.append([asset_id, name, category, purchase_date, purchase_cost])

    date_format = 'yyyy-mm-dd'
    cost_format = '$#,##0.00'

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Asset ID
        ws.cell(row=r, column=2, value=row_data[1])  # Name
        ws.cell(row=r, column=3, value=row_data[2])  # Category
        d_cell = ws.cell(row=r, column=4, value=row_data[3])  # Purchase Date
        d_cell.number_format = date_format
        c_cell = ws.cell(row=r, column=5, value=row_data[4])  # Purchase Cost
        c_cell.number_format = cost_format
        # Column F (Age) intentionally left EMPTY

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
