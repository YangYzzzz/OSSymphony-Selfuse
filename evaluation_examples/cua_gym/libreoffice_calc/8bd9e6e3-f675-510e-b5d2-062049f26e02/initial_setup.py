"""
Initial Setup: Create a spreadsheet with study hours vs exam scores data
Task ID: calc_chart_scatter_regression_080
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_scatter_regression_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: StudyData ---
    ws = wb.active
    ws.title = 'StudyData'

    # Headers
    ws['A1'] = 'Study Hours'
    ws['B1'] = 'Exam Score'

    # Data rows as specified in context
    data = [
        (1,   52),
        (2,   58),
        (2.5, 63),
        (3,   67),
        (4,   72),
        (4.5, 76),
        (5,   81),
        (6,   85),
        (7,   88),
        (8,   92),
    ]

    for r, (hours, score) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=hours)
        ws.cell(row=r, column=2, value=score)

    # No charts in initial file — agent must create the chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
