"""
Reward Script: Configure print setup for 'Summary' sheet — landscape A4 with 1.5cm margins
Task ID: calc_gg1_004
Domain: libreoffice_calc
Scoring:
  Component 1: Orientation set to landscape (0.3 pts)
  Component 2: Paper size set to A4 (0.3 pts)
  Component 3: All four margins set to 1.5 cm (0.4 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_004'

# 1.5 cm in inches = 1.5 / 2.54
EXPECTED_MARGIN_INCHES = 1.5 / 2.54  # ~0.5905511811023622
MARGIN_TOLERANCE = 0.02  # allow small float rounding tolerance


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Summary' sheet must exist
    if 'Summary' not in wb.sheetnames:
        print(f"FAIL: 'Summary' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']
    ps = ws.page_setup
    pm = ws.page_margins

    # Component 1: Orientation set to landscape (0.3 points)
    try:
        orientation = ps.orientation
        if orientation == 'landscape':
            print(f"PASS: Component 1 — Orientation is landscape (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Expected orientation 'landscape', found '{orientation}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Paper size set to A4 (paperSize=9) (0.3 points)
    try:
        paper_size = ps.paperSize
        if paper_size == 9:
            print(f"PASS: Component 2 — Paper size is A4 (paperSize=9) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Expected paperSize 9 (A4), found {paper_size}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All four margins set to 1.5 cm (~0.5906 inches) (0.4 points)
    try:
        margins = {
            'top': pm.top,
            'bottom': pm.bottom,
            'left': pm.left,
            'right': pm.right,
        }
        failed_margins = []
        for name, val in margins.items():
            if val is None or abs(float(val) - EXPECTED_MARGIN_INCHES) > MARGIN_TOLERANCE:
                print(f"FAIL: Component 3 — Margin '{name}' expected ~{EXPECTED_MARGIN_INCHES:.4f} in, found {val}")
                failed_margins.append(name)

        if len(failed_margins) == 0:
            print(f"PASS: Component 3 — All four margins set to 1.5 cm ({EXPECTED_MARGIN_INCHES:.4f} in) (0.4 pts)")
            total_score += 0.4
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
