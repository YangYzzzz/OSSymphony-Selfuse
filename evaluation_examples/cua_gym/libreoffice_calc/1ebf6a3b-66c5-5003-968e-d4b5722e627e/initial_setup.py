"""
Initial Setup: Configure print setup for Summary sheet
Task ID: calc_gg1_004
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Summary Sheet (wide table A-M, 40 rows) ---
    ws = wb.active
    ws.title = 'Summary'

    # Headers (columns A through M)
    headers = [
        'Department', 'Q1 Revenue', 'Q2 Revenue', 'Q3 Revenue', 'Q4 Revenue',
        'Total Revenue', 'Q1 Expenses', 'Q2 Expenses', 'Q3 Expenses', 'Q4 Expenses',
        'Total Expenses', 'Net Profit', 'Margin %'
    ]
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Department data - 39 rows of financial summary data (rows 2-40)
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Human Resources', 'Finance',
        'Operations', 'Customer Support', 'Research & Development', 'Legal',
        'Product Management', 'Quality Assurance', 'IT Infrastructure',
        'Business Development', 'Supply Chain', 'Procurement',
        'Training & Development', 'Corporate Strategy', 'Public Relations',
        'Investor Relations', 'Facilities Management', 'Data Analytics',
        'Compliance', 'Internal Audit', 'Risk Management', 'Treasury',
        'Accounts Payable', 'Accounts Receivable', 'Payroll',
        'Benefits Administration', 'Talent Acquisition', 'Security',
        'Environmental Health', 'Logistics', 'Warehouse Operations',
        'Fleet Management', 'Real Estate', 'Insurance', 'Tax',
        'Mergers & Acquisitions'
    ]

    import random
    random.seed(42)

    currency_fmt = '$#,##0'
    pct_fmt = '0.0%'

    for r, dept in enumerate(departments, 2):
        base_rev = random.randint(150000, 2500000)
        ws.cell(row=r, column=1, value=dept)

        # Q1-Q4 Revenue
        q_revs = []
        for q in range(4):
            val = int(base_rev * (0.8 + random.random() * 0.4))
            q_revs.append(val)
            c = ws.cell(row=r, column=2 + q, value=val)
            c.number_format = currency_fmt

        # Total Revenue (col F = 6)
        total_rev = sum(q_revs)
        c = ws.cell(row=r, column=6, value=total_rev)
        c.number_format = currency_fmt

        # Q1-Q4 Expenses
        q_exps = []
        for q in range(4):
            val = int(q_revs[q] * (0.5 + random.random() * 0.35))
            q_exps.append(val)
            c = ws.cell(row=r, column=7 + q, value=val)
            c.number_format = currency_fmt

        # Total Expenses (col K = 11)
        total_exp = sum(q_exps)
        c = ws.cell(row=r, column=11, value=total_exp)
        c.number_format = currency_fmt

        # Net Profit (col L = 12)
        net = total_rev - total_exp
        c = ws.cell(row=r, column=12, value=net)
        c.number_format = currency_fmt

        # Margin % (col M = 13)
        margin = net / total_rev if total_rev != 0 else 0
        c = ws.cell(row=r, column=13, value=round(margin, 4))
        c.number_format = pct_fmt

    # Set column widths for readability
    ws.column_dimensions['A'].width = 28
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col_letter].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Page setup: Portrait orientation on Letter paper (default state)
    # These are the defaults we want - portrait, letter, standard margins
    # openpyxl uses the worksheet.page_setup for print settings
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
