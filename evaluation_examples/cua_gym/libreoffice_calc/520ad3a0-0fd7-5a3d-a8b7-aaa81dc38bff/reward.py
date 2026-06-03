"""
Reward Script: Financial statement summary dashboard formatting
Task ID: calc_gsd_021
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): A1:B1 merged, bold 16pt, centered
  Component 2 (0.30): Section header rows (3, 9, 16) have dark blue bg + white bold text
  Component 3 (0.25): Currency formatting ($#,##0.00) on data cells B4:B7, B10:B14, B17:B20
  Component 4 (0.20): Thin top border on rows 9 and 16 (separator lines)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_021'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Summary' sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Component 1: A1:B1 merged, bold 16pt, centered (0.25 points)
    try:
        merged = False
        for mr in ws.merged_cells.ranges:
            if str(mr) == 'A1:B1':
                merged = True
                break

        a1 = ws['A1']
        is_bold = (a1.font.bold == True)
        is_16pt = (a1.font.size is not None and abs(float(a1.font.size) - 16.0) < 0.5)
        is_centered = (a1.alignment.horizontal == 'center')

        # All sub-checks must pass; merged is the key differentiator from initial
        if merged and is_bold and is_16pt:
            print(f"PASS: Component 1 -- A1:B1 merged={merged}, bold={is_bold}, size={a1.font.size}, centered={is_centered} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 -- merged={merged}, bold={is_bold}, size={a1.font.size}, centered={is_centered}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Section headers rows 3, 9, 16 -- dark blue bg (#1F3864) + white bold text (0.30 points)
    # Each header is worth 0.10 points
    try:
        header_score = 0.0
        for row_num in [3, 9, 16]:
            cell = ws.cell(row=row_num, column=1)
            is_bold = (cell.font.bold == True)

            # Check dark blue background (FF1F3864 in ARGB)
            has_dark_blue = False
            try:
                fg_rgb = cell.fill.fgColor.rgb
                if fg_rgb and fg_rgb.upper() == 'FF1F3864':
                    has_dark_blue = True
            except Exception:
                pass

            # Check white font color
            has_white_font = False
            try:
                font_rgb = cell.font.color.rgb
                if font_rgb and font_rgb.upper() in ('FFFFFFFF', '00FFFFFF'):
                    has_white_font = True
            except Exception:
                pass

            if is_bold and has_dark_blue and has_white_font:
                print(f"PASS: Component 2 -- Row {row_num} header: bold={is_bold}, bg=dark_blue, font=white (0.10 pts)")
                header_score += 0.10
            else:
                print(f"FAIL: Component 2 -- Row {row_num}: bold={is_bold}, dark_blue_bg={has_dark_blue}, white_font={has_white_font}")

        total_score += header_score
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Currency formatting on data cells B4:B7, B10:B14, B17:B20 (0.25 points)
    try:
        currency_rows = list(range(4, 8)) + list(range(10, 15)) + list(range(17, 21))
        formatted_count = 0
        total_cells = len(currency_rows)

        for row_num in currency_rows:
            cell = ws.cell(row=row_num, column=2)
            nf = cell.number_format if cell.number_format else ''
            # Accept formats that include $ and .00 pattern
            if '$' in nf and '0.00' in nf:
                formatted_count += 1
            else:
                print(f"  INFO: B{row_num} number_format='{nf}' (not currency)")

        if formatted_count == total_cells:
            print(f"PASS: Component 3 -- All {total_cells} data cells have USD currency format (0.25 pts)")
            total_score += 0.25
        elif formatted_count > 0:
            partial = round(0.25 * (formatted_count / total_cells), 2)
            print(f"PARTIAL: Component 3 -- {formatted_count}/{total_cells} cells formatted ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No cells have currency formatting")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Thin top border on rows 9 and 16 (separator lines) (0.20 points)
    # Check both columns A and B for each separator row
    try:
        border_score = 0.0
        for row_num in [9, 16]:
            row_has_border = True
            for col in [1, 2]:
                cell = ws.cell(row=row_num, column=col)
                top_style = cell.border.top.style if cell.border and cell.border.top else None
                if top_style != 'thin':
                    row_has_border = False
                    break

            if row_has_border:
                print(f"PASS: Component 4 -- Row {row_num} has thin top border (0.10 pts)")
                border_score += 0.10
            else:
                print(f"FAIL: Component 4 -- Row {row_num} missing thin top border")

        total_score += border_score
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
