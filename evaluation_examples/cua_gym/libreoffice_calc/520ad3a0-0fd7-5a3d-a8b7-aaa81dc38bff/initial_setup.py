"""
Initial Setup: Financial statement summary dashboard - unformatted data
Task ID: calc_gsd_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Summary"

    # --- Title row (plain text, no formatting) ---
    ws.cell(row=1, column=1, value="FINANCIAL SUMMARY FY2024")

    # --- REVENUE section ---
    ws.cell(row=3, column=1, value="REVENUE")
    revenue_data = [
        ("Product Sales", 2847500.00),
        ("Service Revenue", 1235000.00),
        ("Licensing", 487250.00),
        ("Other", 93400.00),
    ]
    for i, (label, amount) in enumerate(revenue_data, 4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=amount)

    # Row 8 intentionally blank (separator space)

    # --- COSTS section ---
    ws.cell(row=9, column=1, value="COSTS")
    costs_data = [
        ("COGS", 1423750.00),
        ("R&D", 568000.00),
        ("Sales & Marketing", 412300.00),
        ("G&A", 287600.00),
        ("Depreciation", 145800.00),
    ]
    for i, (label, amount) in enumerate(costs_data, 10):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=amount)

    # Row 15 intentionally blank (separator space)

    # --- PROFITABILITY section ---
    ws.cell(row=16, column=1, value="PROFITABILITY")
    profit_data = [
        ("Gross Profit", 1423750.00),
        ("EBITDA", 1243400.00),
        ("Operating Income", 1097600.00),
        ("Net Income", 825350.00),
    ]
    for i, (label, amount) in enumerate(profit_data, 17):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=amount)

    # Set reasonable column widths for readability
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
