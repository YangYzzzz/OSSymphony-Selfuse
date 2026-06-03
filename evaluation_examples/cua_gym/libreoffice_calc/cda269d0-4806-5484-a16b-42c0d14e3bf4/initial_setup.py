"""
Initial Setup: Spreadsheet with freeze panes at C5
Task ID: calc_tbl_070
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_070'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Regional Sales ---
    ws1 = wb.active
    ws1.title = 'Regional Sales'

    # Headers (row 1)
    headers = ['Region', 'Product', 'Q1 Revenue', 'Q2 Revenue', 'Q3 Revenue', 'Q4 Revenue', 'Annual Total', 'Growth %']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # Sub-headers (rows 2-4) representing category groupings
    category_rows = [
        ['North America', '', '', '', '', '', '', ''],
        ['  United States', '', '', '', '', '', '', ''],
        ['  Canada', '', '', '', '', '', '', ''],
    ]
    cat_font = Font(name='Calibri', size=11, bold=True)
    for r, row_data in enumerate(category_rows, 2):
        cell = ws1.cell(row=r, column=1, value=row_data[0])
        cell.font = cat_font

    # Data rows (5-24) with realistic regional sales data
    data = [
        ['  Toronto Office', 'Cloud Suite', 145200, 162300, 178400, 195600, None, None],
        ['  Toronto Office', 'Data Platform', 89300, 95600, 102100, 110800, None, None],
        ['  Vancouver Office', 'Cloud Suite', 67800, 72400, 78900, 84200, None, None],
        ['  Vancouver Office', 'Data Platform', 45100, 48700, 52300, 56800, None, None],
        ['  New York Office', 'Cloud Suite', 234500, 256800, 278900, 301200, None, None],
        ['  New York Office', 'Data Platform', 178900, 192300, 208700, 224500, None, None],
        ['  San Francisco Office', 'Cloud Suite', 312400, 338700, 365200, 392800, None, None],
        ['  San Francisco Office', 'Data Platform', 198700, 215400, 232800, 251300, None, None],
        ['  Chicago Office', 'Cloud Suite', 123400, 134800, 146200, 158700, None, None],
        ['  Chicago Office', 'Data Platform', 87600, 94200, 101800, 109500, None, None],
        ['Europe', '', '', '', '', '', '', ''],
        ['  London Office', 'Cloud Suite', 189300, 205600, 222800, 241200, None, None],
        ['  London Office', 'Data Platform', 134200, 145800, 158300, 171600, None, None],
        ['  Berlin Office', 'Cloud Suite', 98700, 107200, 116400, 126300, None, None],
        ['  Berlin Office', 'Data Platform', 67400, 73200, 79800, 86900, None, None],
        ['  Paris Office', 'Cloud Suite', 112300, 121800, 132400, 143600, None, None],
        ['  Paris Office', 'Data Platform', 78900, 85600, 93200, 101400, None, None],
        ['Asia Pacific', '', '', '', '', '', '', ''],
        ['  Tokyo Office', 'Cloud Suite', 156800, 170200, 184900, 200300, None, None],
        ['  Tokyo Office', 'Data Platform', 109400, 118700, 128900, 139800, None, None],
    ]

    for r, row_data in enumerate(data, 5):
        for c, val in enumerate(row_data, 1):
            if val is not None:
                cell = ws1.cell(row=r, column=c, value=val)
                cell.border = border
                if c >= 3 and c <= 6 and isinstance(val, (int, float)):
                    cell.number_format = '#,##0'

    # Set column widths
    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 16
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws1.column_dimensions[col_letter].width = 14

    # Set row 1 height
    ws1.row_dimensions[1].height = 25

    # FREEZE PANES at C5 (rows 1-4 and columns A-B frozen)
    ws1.freeze_panes = 'C5'

    # --- Sheet 2: Product Summary ---
    ws2 = wb.create_sheet('Product Summary')
    prod_headers = ['Product', 'Total Units Sold', 'Avg Unit Price', 'Total Revenue', 'Market Share %']
    for col, h in enumerate(prod_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = border

    prod_data = [
        ['Cloud Suite', 4520, 285.50, 1290660, 58.2],
        ['Data Platform', 3180, 195.75, 622485, 28.1],
        ['Analytics Pro', 1240, 342.00, 424080, 19.1],
        ['Security Shield', 890, 178.25, 158643, 7.2],
        ['DevOps Toolkit', 1560, 225.00, 351000, 15.8],
    ]
    for r, row_data in enumerate(prod_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = border
            if c in [3, 4]:
                cell.number_format = '#,##0.00'
            elif c == 5:
                cell.number_format = '0.0'

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 18
    ws2.column_dimensions['C'].width = 15
    ws2.column_dimensions['D'].width = 15
    ws2.column_dimensions['E'].width = 15

    # --- Sheet 3: Quarterly Targets ---
    ws3 = wb.create_sheet('Quarterly Targets')
    target_headers = ['Quarter', 'Target Revenue', 'Actual Revenue', 'Variance', 'Status']
    for col, h in enumerate(target_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.border = border

    target_data = [
        ['Q1 2025', 2500000, 2478200, -21800, 'Near Target'],
        ['Q2 2025', 2750000, 2812400, 62400, 'Exceeded'],
        ['Q3 2025', 3000000, 3145600, 145600, 'Exceeded'],
        ['Q4 2025', 3250000, 3418500, 168500, 'Exceeded'],
    ]
    for r, row_data in enumerate(target_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = border
            if c in [2, 3, 4]:
                cell.number_format = '#,##0'

    ws3.column_dimensions['A'].width = 14
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 18
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
