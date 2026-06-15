"""
Initial Setup: Marketing channel performance spreadsheet (no total row, no chart)
Task ID: osworld_calc_total_row_line_chart_008
Domain: libreoffice_calc

Creates a spreadsheet with marketing channel data (8 weeks, multiple channels)
WITHOUT a total row or any chart — those are what the agent must add.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marketing Performance"

    # --- Headers ---
    headers = ["Channel", "Week 1", "Week 2", "Week 3", "Week 4",
               "Week 5", "Week 6", "Week 7", "Week 8"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name="Calibri", size=11)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, name="Calibri", size=11, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # --- Marketing channel data (realistic, 8 channels, 8 weeks) ---
    data = [
        ["Email Campaigns",    12450, 13820, 11930, 14560, 15200, 13750, 14890, 16320],
        ["Social Media Ads",   18900, 20340, 19780, 22100, 21450, 23890, 22670, 24100],
        ["Search (SEM/PPC)",   34500, 36200, 35100, 38400, 37800, 40100, 39200, 42300],
        ["Organic Search",     28700, 29500, 30200, 31800, 30900, 32500, 33100, 34800],
        ["Affiliate Network",   8200,  7950,  8450,  9100,  8750,  9400,  9800, 10200],
        ["Display Ads",        11300, 12100, 11800, 13200, 12900, 13800, 14100, 15200],
        ["Influencer Collab",   6800,  7200,  7500,  8100,  7800,  8400,  8900,  9500],
        ["Content Marketing",   9400,  9800, 10200, 10700, 10500, 11200, 11600, 12300],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0'

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
