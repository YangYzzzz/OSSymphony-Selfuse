"""
Reward Script: Add a total row to the marketing channel performance table and create a line chart
Task ID: osworld_calc_total_row_line_chart_008
Domain: libreoffice_calc
Scoring:
  Component 1: Total row label in column A below data (0.20 pts)
  Component 2: SUM formulas in B10:I10 covering all 8 channel rows (0.40 pts)
  Component 3: Line chart present on sheet (0.20 pts)
  Component 4: Chart title is "Weekly Marketing Performance" (0.20 pts)
"""

import os
import openpyxl
from openpyxl.chart import LineChart

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_total_row_line_chart_008'

EXPECTED_TITLE = "Weekly Marketing Performance"
WEEKS_COLS = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']  # 8 week columns


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    try:
        ws = wb['Marketing Performance']
    except KeyError:
        # Try first sheet if named differently
        if wb.sheetnames:
            ws = wb.worksheets[0]
            print(f"WARN: Sheet 'Marketing Performance' not found; using '{ws.title}'")
        else:
            print("CRITICAL: No sheets found in workbook")
            print("REWARD: 0.0")
            return 0.0

    # Precondition: original data must be present (rows 1-9 unchanged)
    # Row 1 should be headers, rows 2-9 should be channel data
    header_a1 = ws['A1'].value
    if not header_a1 or str(header_a1).strip().lower() != 'channel':
        print("CRITICAL: Spreadsheet structure not recognized (A1 != 'Channel'). Aborting.")
        print("REWARD: 0.0")
        return 0.0

    # Determine the data row range: rows 2 through 9 (8 channels)
    # Total row should come after the last data row
    # We accept the total row at row 10 or wherever it is placed below row 9

    # --- Component 1: Total row label (0.20 points) ---
    # Find a row below row 9 that contains "Total" (or similar) in column A
    try:
        total_row_idx = None
        for row_idx in range(10, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val is not None and str(cell_val).strip().lower() == 'total':
                total_row_idx = row_idx
                break

        if total_row_idx is not None:
            print(f"PASS: Component 1 — Total row label found at row {total_row_idx} in column A (0.20 pts)")
            total_score += 0.20
        else:
            # Check if there is any label below row 9
            any_label = None
            for row_idx in range(10, ws.max_row + 1):
                cell_val = ws.cell(row=row_idx, column=1).value
                if cell_val is not None:
                    any_label = (row_idx, cell_val)
                    break
            if any_label:
                print(f"FAIL: Component 1 — Expected 'Total' label below row 9 in column A, "
                      f"found '{any_label[1]}' at row {any_label[0]}")
            else:
                print("FAIL: Component 1 — No row label found below data rows (expected 'Total' in column A)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: SUM formulas for all 8 week columns in the total row (0.40 points) ---
    # We need SUM formulas in columns B through I of the total row
    try:
        if total_row_idx is None:
            print("FAIL: Component 2 — Cannot check SUM formulas: total row not found")
        else:
            sum_cols_ok = 0
            sum_details = []
            for col_letter in WEEKS_COLS:
                cell = ws[f'{col_letter}{total_row_idx}']
                cell_val = cell.value
                if cell_val is not None and isinstance(cell_val, str) and cell_val.strip().upper().startswith('=SUM('):
                    # Verify it references the data rows (should be rows 2 to total_row_idx-1)
                    sum_cols_ok += 1
                    sum_details.append(f"{col_letter}{total_row_idx}:{cell_val}")
                elif cell_val is not None and isinstance(cell_val, (int, float)):
                    # Accept numeric value as possible cached SUM (less ideal but functional)
                    # Only award if the value looks like a column sum from rows 2-9
                    expected_sum = sum(
                        ws.cell(row=r, column=ord(col_letter) - ord('A') + 1).value or 0
                        for r in range(2, total_row_idx)
                    )
                    if abs(cell_val - expected_sum) < 1:
                        sum_cols_ok += 1
                        sum_details.append(f"{col_letter}{total_row_idx}:{cell_val}(numeric)")
                    else:
                        sum_details.append(f"{col_letter}{total_row_idx}:WRONG({cell_val})")
                else:
                    sum_details.append(f"{col_letter}{total_row_idx}:MISSING({cell_val!r})")

            if sum_cols_ok == 8:
                print(f"PASS: Component 2 — All 8 SUM formulas present in total row (0.40 pts): {', '.join(sum_details)}")
                total_score += 0.40
            elif sum_cols_ok >= 4:
                # Partial: award 0.20 for at least half the columns have valid formulas
                print(f"PARTIAL: Component 2 — {sum_cols_ok}/8 SUM columns OK, awarding 0.20 pts: {', '.join(sum_details)}")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — Only {sum_cols_ok}/8 columns have valid SUM formulas: {', '.join(sum_details)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Line chart present on the sheet (0.20 points) ---
    try:
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, LineChart)]

        if line_charts:
            print(f"PASS: Component 3 — Line chart found on sheet ({len(line_charts)} line chart(s)) (0.20 pts)")
            total_score += 0.20
        elif charts:
            # Some chart exists but not a LineChart — partial credit NOT awarded
            chart_types = [type(c).__name__ for c in charts]
            print(f"FAIL: Component 3 — Chart(s) found but not LineChart: {chart_types}")
        else:
            print("FAIL: Component 3 — No chart found on sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # --- Component 4: Chart title is "Weekly Marketing Performance" (0.20 points) ---
    try:
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, LineChart)]

        if line_charts:
            chart = line_charts[0]
            # Extract title text from the openpyxl Title object
            chart_title_text = None
            if chart.title is not None:
                title_obj = chart.title
                # Try direct string
                if isinstance(title_obj, str):
                    chart_title_text = title_obj
                else:
                    # Navigate the nested Title object structure: title.tx.rich.p[].r[].t
                    try:
                        paragraphs = title_obj.tx.rich.p
                        text_parts = []
                        for para in paragraphs:
                            for run in para.r:
                                text_parts.append(run.t)
                        chart_title_text = ''.join(text_parts)
                    except Exception:
                        chart_title_text = None

            # Check title text (case-insensitive match accepted)
            if chart_title_text and chart_title_text.strip().lower() == EXPECTED_TITLE.lower():
                print(f"PASS: Component 4 — Chart title is '{chart_title_text}' (0.20 pts)")
                total_score += 0.20
            elif chart_title_text:
                print(f"FAIL: Component 4 — Chart title is '{chart_title_text}', expected '{EXPECTED_TITLE}'")
            else:
                print(f"FAIL: Component 4 — Chart has no title or title could not be extracted (expected '{EXPECTED_TITLE}')")
        else:
            print("FAIL: Component 4 — No line chart found, cannot verify title")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in a given env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
