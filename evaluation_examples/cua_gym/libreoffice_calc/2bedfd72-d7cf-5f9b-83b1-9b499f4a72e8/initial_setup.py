"""
Initial Setup: Cross-check PDF invoices against bank statement spreadsheet
Task ID: osworld_multi_apps_doc_pdf_calc_007
Domain: libreoffice_calc (multi-app: PDF, Calc, Writer)

Creates:
- /home/user/Desktop/pdf_invoices/ folder with 4 PDF invoice files
- /home/user/Desktop/bank_statement.ods with Sheet1 (Transactions)
  containing transactions - Vendors A and C present, B and D missing
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_doc_pdf_calc_007'
BANK_STATEMENT = f'{DESKTOP}/bank_statement.ods'
PDF_FOLDER = f'{DESKTOP}/pdf_invoices'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_pdf_folder():
    """Create pdf_invoices directory on Desktop."""
    os.makedirs(PDF_FOLDER, exist_ok=True)
    print(f'PDF folder created: {PDF_FOLDER}')


def create_invoice_pdf(filename, vendor_name, amount, date_str, ref_num, item_desc):
    """Create a simple PDF invoice using fpdf2."""
    try:
        from fpdf import FPDF
    except ImportError:
        subprocess.run(['pip3', 'install', 'fpdf2'], check=True)
        from fpdf import FPDF

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font('Helvetica', 'B', 20)
    pdf.cell(0, 12, 'INVOICE', new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(4)

    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 8, f'{vendor_name}', new_x='LMARGIN', new_y='NEXT', align='C')
    pdf.ln(6)

    pdf.set_font('Helvetica', '', 11)
    pdf.cell(50, 8, 'Reference Number:', border=0)
    pdf.cell(0, 8, ref_num, new_x='LMARGIN', new_y='NEXT')

    pdf.cell(50, 8, 'Invoice Date:', border=0)
    pdf.cell(0, 8, date_str, new_x='LMARGIN', new_y='NEXT')

    pdf.cell(50, 8, 'Description:', border=0)
    pdf.cell(0, 8, item_desc, new_x='LMARGIN', new_y='NEXT')
    pdf.ln(6)

    pdf.set_font('Helvetica', 'B', 12)
    pdf.cell(50, 10, 'AMOUNT DUE:', border=0)
    pdf.set_font('Helvetica', 'B', 14)
    pdf.cell(0, 10, f'${amount:.2f}', new_x='LMARGIN', new_y='NEXT')
    pdf.ln(8)

    pdf.set_font('Helvetica', 'I', 9)
    pdf.cell(0, 6, 'Please remit payment within 30 days.', new_x='LMARGIN', new_y='NEXT', align='C')

    out_path = os.path.join(PDF_FOLDER, filename)
    pdf.output(out_path)
    print(f'  PDF invoice created: {out_path}')


def create_bank_statement():
    """Create bank_statement.ods with Sheet1 Transactions.
    Vendors A and C are recorded; Vendors B and D are MISSING from bank statement.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        subprocess.run(['pip3', 'install', 'openpyxl'], check=True)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions'

    # Headers
    headers = ['Date', 'Reference', 'Vendor', 'Description', 'Amount', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Transaction data - Vendors A and C are present; B and D are missing
    # Adding realistic business transactions for February 2025
    transactions = [
        # (Date, Reference, Vendor, Description, Amount, Status)
        ('2025-02-03', 'INT-2025-001', 'Office Depot',      'Office supplies - Q1',           485.30, 'Cleared'),
        ('2025-02-05', 'VA-001',       'Vendor A Solutions', 'Software license renewal',      1250.00, 'Cleared'),
        ('2025-02-06', 'PAY-2025-002', 'City Utilities',    'Monthly utility bill',            312.75, 'Cleared'),
        ('2025-02-07', 'INT-2025-003', 'FedEx',             'Shipping charges - Jan',          98.40, 'Cleared'),
        ('2025-02-10', 'VC-003',       'Vendor C Corp',     'IT consulting services',          340.00, 'Cleared'),
        ('2025-02-11', 'PAY-2025-005', 'Amazon Web Services','Cloud hosting - Feb',           1875.20, 'Cleared'),
        ('2025-02-12', 'INT-2025-006', 'Staples',           'Printer cartridges',              127.50, 'Cleared'),
        ('2025-02-13', 'PAY-2025-007', 'Delta Airlines',    'Travel - Sales conference',       623.80, 'Cleared'),
        ('2025-02-17', 'PAY-2025-008', 'Verizon',          'Phone & data plan - Feb',          245.00, 'Cleared'),
        ('2025-02-19', 'INT-2025-009', 'UPS',              'Courier services',                  67.20, 'Cleared'),
        ('2025-02-20', 'PAY-2025-010', 'Adobe Inc.',       'Creative Cloud subscription',      599.88, 'Cleared'),
        ('2025-02-24', 'INT-2025-011', 'Costco',           'Office kitchen supplies',          143.60, 'Cleared'),
        ('2025-02-26', 'PAY-2025-012', 'Zoom Video',       'Video conferencing - monthly',     149.90, 'Cleared'),
        ('2025-02-28', 'INT-2025-013', 'Comcast Business', 'Internet service - Feb',           189.99, 'Cleared'),
    ]

    for row_idx, (date_val, ref, vendor, desc, amount, status) in enumerate(transactions, 2):
        ws.cell(row=row_idx, column=1, value=date_val)
        ws.cell(row=row_idx, column=2, value=ref)
        ws.cell(row=row_idx, column=3, value=vendor)
        ws.cell(row=row_idx, column=4, value=desc)
        cell_amount = ws.cell(row=row_idx, column=5, value=amount)
        cell_amount.number_format = '$#,##0.00'
        ws.cell(row=row_idx, column=6, value=status)

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 36
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save as xlsx first, then we need to save as ODS
    # Note: openpyxl can save as .ods via odfpy
    wb.save(BANK_STATEMENT)
    print(f'Bank statement created: {BANK_STATEMENT}')


def create_initial():
    """Main function to set up initial environment."""
    # Create directories
    os.makedirs(DESKTOP, exist_ok=True)
    create_pdf_folder()

    # Create 4 PDF invoices
    print('Creating PDF invoices...')
    create_invoice_pdf(
        'vendor_a.pdf',
        'Vendor A Solutions',
        1250.00,
        '2025-02-05',
        'VA-001',
        'Software license renewal - Annual subscription'
    )
    create_invoice_pdf(
        'vendor_b.pdf',
        'Vendor B Technologies',
        875.50,
        '2025-02-08',
        'VB-002',
        'Hardware maintenance contract - Q1'
    )
    create_invoice_pdf(
        'vendor_c.pdf',
        'Vendor C Corp',
        340.00,
        '2025-02-10',
        'VC-003',
        'IT consulting services - February'
    )
    create_invoice_pdf(
        'vendor_d.pdf',
        'Vendor D Logistics',
        2100.00,
        '2025-02-14',
        'VD-004',
        'Warehousing and distribution services - Feb'
    )
    print('All PDF invoices created.')

    # Create bank statement spreadsheet
    print('Creating bank statement...')
    create_bank_statement()

    # GUI-ready startup: open the bank statement in LibreOffice Calc
    # and open the pdf_invoices folder in file manager
    print('Launching GUI applications...')
    launch_gui(f'libreoffice --calc "{BANK_STATEMENT}"', delay_sec=3.0)
    launch_gui(f'nautilus "{PDF_FOLDER}"', delay_sec=1.5)

    print('GUI_READY: Launched LibreOffice Calc and file manager with DISPLAY=:0')
    print(f'Initial setup complete for task: {TASK_ID}')


create_initial()
