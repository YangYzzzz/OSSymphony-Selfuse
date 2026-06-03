"""
Initial Setup: ICLR 2024 spotlight papers spreadsheet
Task ID: osworld_multi_apps_web_papers_009
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Creates iclr_spotlights.ods on Desktop with existing non-transformer spotlight entries.
The agent must find transformer/attention spotlight papers from ICLR 2024
and append them to this file.
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user/Desktop'
TASK_ID = 'iclr_spotlights'
OUTPUT = f'{WORKDIR}/{TASK_ID}.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties, TableCellProperties, TableColumnProperties
        from odf import style as odfstyle

        doc = OpenDocumentSpreadsheet()

        # Create a bold style for headers
        header_style = Style(name="HeaderStyle", family="table-cell")
        header_style.addElement(TextProperties(fontweight="bold"))
        doc.styles.addElement(header_style)

        table = Table(name="ICLR Spotlights")

        def make_cell(value, stylename=None):
            if stylename:
                tc = TableCell(stylename=stylename, valuetype="string")
            else:
                tc = TableCell(valuetype="string")
            tc.addElement(P(text=str(value)))
            return tc

        def make_number_cell(value):
            # Use string type for compatibility - rating will be stored as text
            tc = TableCell(valuetype="string")
            tc.addElement(P(text=str(value)))
            return tc

        # Header row
        header_row = TableRow()
        for col in ["Title", "Authors", "Rating", "Track", "Notes"]:
            header_row.addElement(make_cell(col, stylename="HeaderStyle"))
        table.addElement(header_row)

        # Pre-existing non-transformer spotlight entries
        existing_data = [
            [
                "Vision Mamba: Efficient Visual Representation Learning with Bidirectional State Space Model",
                "Lianghui Zhu, Bencheng Liao, Qian Zhang et al.",
                8,
                "Spotlight",
                ""
            ],
            [
                "Diffusion-based Image Translation with Label Guidance for Domain Adaptive Semantic Segmentation",
                "Duo Peng, Ping Hu, Xinzhe Luo et al.",
                6,
                "Spotlight",
                ""
            ],
            [
                "FedAvg with Fine Tuning: Local Updates Lead to Representation Learning",
                "Liam Collins, Hamed Hassani, Aryan Mokhtari et al.",
                6,
                "Spotlight",
                ""
            ],
            [
                "Flow Matching on General Geometries",
                "Ricky T. Q. Chen, Yaron Lipman",
                8,
                "Spotlight",
                ""
            ],
            [
                "LLM-Planner: Few-Shot Grounded Planning for Embodied Agents with Large Language Models",
                "Chan Hee Song, Jiaman Wu, Clayton Washington et al.",
                6,
                "Spotlight",
                ""
            ],
            [
                "Contrastive Preference Learning: Learning from Human Feedback without RL",
                "Joey Hejna, Rafael Rafailov, Harshit Sikchi et al.",
                6,
                "Spotlight",
                ""
            ],
            [
                "Generalization in diffusion models arises from geometry-adaptive harmonic representations",
                "Zahra Kadkhodaie, Florentin Guth, Eero Simoncelli et al.",
                8,
                "Spotlight",
                ""
            ],
        ]

        for row_data in existing_data:
            row = TableRow()
            for i, val in enumerate(row_data):
                if i == 2:  # Rating column (numeric)
                    row.addElement(make_number_cell(val))
                else:
                    row.addElement(make_cell(val))
            table.addElement(row)

        doc.spreadsheet.addElement(table)
        doc.save(OUTPUT)
        print(f'Initial file created: {OUTPUT}')

    except ImportError as e:
        print(f"odfpy not available, falling back to openpyxl: {e}")
        # Fallback: create xlsx and note the issue
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ICLR Spotlights"

        headers = ["Title", "Authors", "Rating", "Track", "Notes"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        existing_data = [
            ["Vision Mamba: Efficient Visual Representation Learning with Bidirectional State Space Model",
             "Lianghui Zhu, Bencheng Liao, Qian Zhang et al.", 8, "Spotlight", ""],
            ["Diffusion-based Image Translation with Label Guidance for Domain Adaptive Semantic Segmentation",
             "Duo Peng, Ping Hu, Xinzhe Luo et al.", 6, "Spotlight", ""],
            ["FedAvg with Fine Tuning: Local Updates Lead to Representation Learning",
             "Liam Collins, Hamed Hassani, Aryan Mokhtari et al.", 6, "Spotlight", ""],
            ["Flow Matching on General Geometries",
             "Ricky T. Q. Chen, Yaron Lipman", 8, "Spotlight", ""],
            ["LLM-Planner: Few-Shot Grounded Planning for Embodied Agents with Large Language Models",
             "Chan Hee Song, Jiaman Wu, Clayton Washington et al.", 6, "Spotlight", ""],
            ["Contrastive Preference Learning: Learning from Human Feedback without RL",
             "Joey Hejna, Rafael Rafailov, Harshit Sikchi et al.", 6, "Spotlight", ""],
            ["Generalization in diffusion models arises from geometry-adaptive harmonic representations",
             "Zahra Kadkhodaie, Florentin Guth, Eero Simoncelli et al.", 8, "Spotlight", ""],
        ]
        for r, row_data in enumerate(existing_data, 2):
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val)

        ods_path = OUTPUT  # save as xlsx first since odfpy failed
        wb.save(ods_path.replace('.ods', '.xlsx'))
        print(f'Fallback xlsx created: {ods_path.replace(".ods", ".xlsx")}')

    # GUI-ready startup: open the ods file in LibreOffice Calc + open Chrome on the ICLR page
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui(
        'google-chrome "https://iclr.cc/virtual/2024/papers.html?filter=titles"',
        delay_sec=2.0
    )
    print('GUI_READY: launched LibreOffice Calc and Chrome with ICLR 2024 papers page')


create_initial()
