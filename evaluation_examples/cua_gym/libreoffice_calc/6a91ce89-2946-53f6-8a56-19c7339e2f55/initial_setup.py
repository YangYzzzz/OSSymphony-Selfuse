"""
Initial Setup: Create a spreadsheet with market analysis data (no charts)
Task ID: calc_chart_scatter_bubble_064
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_scatter_bubble_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: MarketAnalysis ---
    ws = wb.active
    ws.title = 'MarketAnalysis'

    # Headers as specified in task context
    headers = ['Segment', 'Market Size ($B)', 'Growth Rate %', 'Our Revenue ($M)']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data as specified in task context (exactly matching)
    data = [
        ['Enterprise', 45, 8.5, 12.4],
        ['SMB',        28, 15.2, 8.6],
        ['Consumer',   62, 5.8,  4.2],
        ['Government', 18, 3.2,  6.8],
        ['Healthcare', 35, 12.4, 3.1],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20

    # NOTE: No charts in the initial file — the task is to create a bubble chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
