"""
Initial Setup: Add secondary Y-axis to embedded chart
Task ID: calc_gg2_002
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Revenue Sheet ---
    ws = wb.active
    ws.title = "Revenue"

    # Headers
    headers = ["Month", "Revenue (USD)", "Units Sold"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Monthly data for 2025
    data = [
        ["Jan 2025", 185400, 142],
        ["Feb 2025", 197800, 168],
        ["Mar 2025", 224500, 215],
        ["Apr 2025", 256300, 278],
        ["May 2025", 312700, 345],
        ["Jun 2025", 289100, 312],
        ["Jul 2025", 267400, 287],
        ["Aug 2025", 298600, 336],
        ["Sep 2025", 342100, 398],
        ["Oct 2025", 378500, 452],
        ["Nov 2025", 415200, 523],
        ["Dec 2025", 438900, 571],
    ]

    data_font = Font(name="Calibri", size=11)
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c == 2:
                cell.number_format = '$#,##0'
            elif c == 3:
                cell.number_format = '#,##0'
            if c == 1:
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14

    # --- Create Combo Chart (both series on primary axis) ---
    # Use a BarChart for Revenue, then overlay a LineChart for Units Sold
    # Both on PRIMARY axis initially (this is the problem the task asks to fix)

    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Monthly Revenue & Units Sold"
    bar_chart.y_axis.title = "Revenue (USD)"
    bar_chart.x_axis.title = "Month"
    bar_chart.style = 10
    bar_chart.width = 22
    bar_chart.height = 14

    # Revenue data series (bars)
    revenue_data = Reference(ws, min_col=2, min_row=1, max_row=13)
    cats = Reference(ws, min_col=1, min_row=2, max_row=13)
    bar_chart.add_data(revenue_data, titles_from_data=True)
    bar_chart.set_categories(cats)

    # Units Sold data series (line overlaid on same chart, same primary axis)
    line_chart = LineChart()
    units_data = Reference(ws, min_col=3, min_row=1, max_row=13)
    line_chart.add_data(units_data, titles_from_data=True)
    line_chart.series[0].graphicalProperties.line.width = 25000  # thicker line

    # Combine: overlay line on bar chart - both use primary axis
    bar_chart += line_chart

    ws.add_chart(bar_chart, "E2")

    # --- Summary Sheet (extra content for complexity) ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Annual Summary"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True)

    summary_headers = ["Metric", "Value"]
    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    summary_data = [
        ["Total Revenue", "=SUM(Revenue!B2:B13)"],
        ["Total Units Sold", "=SUM(Revenue!C2:C13)"],
        ["Average Monthly Revenue", "=AVERAGE(Revenue!B2:B13)"],
        ["Average Monthly Units", "=AVERAGE(Revenue!C2:C13)"],
        ["Peak Revenue Month", "=INDEX(Revenue!A2:A13,MATCH(MAX(Revenue!B2:B13),Revenue!B2:B13,0))"],
        ["Peak Units Month", "=INDEX(Revenue!A2:A13,MATCH(MAX(Revenue!C2:C13),Revenue!C2:C13,0))"],
    ]
    for r, row_data in enumerate(summary_data, 4):
        ws2.cell(row=r, column=1, value=row_data[0]).font = Font(name="Calibri", size=11)
        cell = ws2.cell(row=r, column=2, value=row_data[1])
        if "Revenue" in row_data[0] and "Month" not in row_data[0]:
            cell.number_format = '$#,##0'

    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
