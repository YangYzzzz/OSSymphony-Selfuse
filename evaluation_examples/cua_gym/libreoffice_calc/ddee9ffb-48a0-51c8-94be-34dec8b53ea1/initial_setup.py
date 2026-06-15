"""
Initial Setup: Invoice spreadsheet with named range 'TaxRate' pointing to G1
Task ID: calc_fmb_named_range_formula_044
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_named_range_formula_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Invoice ---
    ws = wb.active
    ws.title = 'Invoice'

    # Headers in row 1
    headers = ['Invoice #', 'Description', 'Qty', 'Subtotal', 'Tax', 'Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # G1: TaxRate value
    ws['G1'] = 0.0875

    # Row 2: Invoice data — E2 and F2 intentionally empty (task targets E2)
    ws['A2'] = 'INV-001'
    ws['B2'] = 'Consulting Services'
    ws['C2'] = 1
    ws['D2'] = 4500.00
    # E2 = empty (target cell — formula to be added by agent)
    # F2 = empty

    # Additional rows for realistic content
    additional_data = [
        ('INV-002', 'Software Development', 3, 12750.00),
        ('INV-003', 'Project Management', 2, 3200.00),
        ('INV-004', 'UX Design Services', 1, 6800.00),
        ('INV-005', 'Database Optimization', 5, 9250.00),
        ('INV-006', 'Technical Support', 10, 1500.00),
        ('INV-007', 'Cloud Migration', 1, 22000.00),
        ('INV-008', 'Security Audit', 1, 7500.00),
        ('INV-009', 'API Integration', 4, 5600.00),
        ('INV-010', 'Training Workshop', 2, 2400.00),
    ]
    for r, (inv_num, desc, qty, subtotal) in enumerate(additional_data, 3):
        ws[f'A{r}'] = inv_num
        ws[f'B{r}'] = desc
        ws[f'C{r}'] = qty
        ws[f'D{r}'] = subtotal
        # E and F columns left empty for these rows too

    # Define the named range 'TaxRate' pointing to Invoice!$G$1
    defined_name = DefinedName('TaxRate', attr_text="Invoice!$G$1")
    wb.defined_names['TaxRate'] = defined_name

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Named range TaxRate defined pointing to Invoice!$G$1 (value=0.0875)')
    print('E2 is empty — agent must add formula =D2*TaxRate')


create_initial()
