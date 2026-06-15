"""
Reward Script: Use named range 'TaxRate' in a formula in E2 to calculate tax on subtotal in D2
Task ID: calc_fmb_named_range_formula_044
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): E2 contains a formula referencing the 'TaxRate' named range
  Component 2 (0.3): E2 formula also references D2 (correct subtotal cell)
  Component 3 (0.2): No other cells were modified (E3-E11 remain empty/None)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_named_range_formula_044'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if 'Invoice' not in wb.sheetnames:
        print("FAIL: Sheet 'Invoice' not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws = wb['Invoice']

    # Component 1: E2 contains a formula referencing the 'TaxRate' named range (0.5 points)
    # This FAILS on initial (E2 is None) and PASSES on golden (E2 = '=D2*TaxRate')
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str) and e2_value.startswith('='):
            formula_upper = e2_value.upper().replace(' ', '')
            if 'TAXRATE' in formula_upper:
                print(f"PASS: Component 1 — E2 contains a formula referencing TaxRate: {repr(e2_value)} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 1 — E2 has a formula but does not reference TaxRate: {repr(e2_value)}")
        else:
            print(f"FAIL: Component 1 — E2 does not contain a formula, found: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — could not check E2: {e}")

    # Component 2: E2 formula references D2 (the subtotal cell) (0.3 points)
    # This FAILS on initial (E2 is None) and PASSES on golden (=D2*TaxRate references D2)
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str) and e2_value.startswith('='):
            formula_upper = e2_value.upper().replace(' ', '')
            if 'D2' in formula_upper:
                print(f"PASS: Component 2 — E2 formula references D2 (subtotal cell): {repr(e2_value)} (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — E2 formula does not reference D2: {repr(e2_value)}")
        else:
            print(f"FAIL: Component 2 — E2 does not contain a formula, found: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — could not check E2 formula for D2 reference: {e}")

    # Component 3: E2 formula is correct AND no other E-column cells were modified (0.2 points)
    # This verifies correctness of E2 combined with task constraint that only E2 was changed.
    # On initial file, E2 is None so this compound check FAILS.
    # On golden file, E2 has the formula AND E3-E11 are empty, so this PASSES.
    try:
        e2_value = ws['E2'].value
        # First sub-condition: E2 must have a formula (change from initial state)
        e2_is_formula = (e2_value is not None and isinstance(e2_value, str) and e2_value.startswith('='))

        if not e2_is_formula:
            print(f"FAIL: Component 3 — E2 does not have a formula (pre-requisite for purity check): {repr(e2_value)}")
        else:
            # Second sub-condition: No other E-column cells (E3-E11) were modified
            # Collect any non-None values in E3:E11
            modified_cells = [
                f"E{row}={repr(ws.cell(row=row, column=5).value)}"
                for row in range(3, 12)
                if ws.cell(row=row, column=5).value is not None
            ]

            if len(modified_cells) == 0:
                print(f"PASS: Component 3 — E2 has formula and E3:E11 remain empty (task scope respected) (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — E2 has formula but other E-column cells were unexpectedly modified: {modified_cells}")
    except Exception as e:
        print(f"ERROR: Component 3 — could not verify scope purity: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
