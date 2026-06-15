"""
Reward Script: A/B Pricing Experiment Tracker with conversion metrics and confidence intervals
Task ID: calc_gpm_089
Domain: libreoffice_calc
Scoring:
  Component 1: Title merged A1:I1 with text/formatting (0.15)
  Component 2: Test setup data A3:B5 with yellow fill   (0.10)
  Component 3: Headers row 7 with labels and bold        (0.15)
  Component 4: 30 rows of daily data (rows 8-37)         (0.15)
  Component 5: Formulas and number formats in D/G/H      (0.10)
  Component 6: Conditional formatting on H and I columns  (0.10)
  Component 7: Final Results row 39 + summary metrics     (0.10)
  Component 8: Two charts (line + bar) with titles        (0.15)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.chart import LineChart, BarChart

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_089'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet name
    if 'PriceTest' not in wb.sheetnames:
        print("FAIL: Sheet 'PriceTest' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PriceTest']

    # =========================================================================
    # Component 1: Title merged A1:I1 with correct text and formatting (0.15)
    # =========================================================================
    try:
        score_1 = 0.0
        a1_val = ws['A1'].value
        # Check value contains expected text
        if a1_val and 'Pricing Experiment Dashboard' in str(a1_val):
            score_1 += 0.05
            print(f"PASS: C1a — Title text found: {repr(a1_val)}")
        else:
            print(f"FAIL: C1a — Expected 'Pricing Experiment Dashboard' in A1, found: {repr(a1_val)}")

        # Check merge: B1 should be a MergedCell
        if isinstance(ws['B1'], MergedCell):
            score_1 += 0.03
            print("PASS: C1b — A1:I1 merge detected")
        else:
            print("FAIL: C1b — A1:I1 not merged")

        # Check font: bold, size ~14
        font = ws['A1'].font
        if font.bold and font.size and font.size >= 13:
            score_1 += 0.03
            print(f"PASS: C1c — Title font bold={font.bold}, size={font.size}")
        else:
            print(f"FAIL: C1c — Expected bold=True, size>=13, got bold={font.bold}, size={font.size}")

        # Check fill: dark purple (4B0082)
        try:
            fill_rgb = ws['A1'].fill.fgColor.rgb
            if fill_rgb and '4B0082' in str(fill_rgb).upper():
                score_1 += 0.02
                print(f"PASS: C1d — Purple fill: {fill_rgb}")
            else:
                print(f"FAIL: C1d — Expected purple fill (4B0082), found: {fill_rgb}")
        except Exception:
            print("FAIL: C1d — Could not read fill color")

        # Check font color: white
        try:
            font_color = ws['A1'].font.color.rgb
            if font_color and 'FFFFFF' in str(font_color).upper():
                score_1 += 0.02
                print(f"PASS: C1e — White font color: {font_color}")
            else:
                print(f"FAIL: C1e — Expected white font, found: {font_color}")
        except Exception:
            print("FAIL: C1e — Could not read font color")

        total_score += score_1
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Test setup data A3:B5 with yellow fill (0.10)
    # =========================================================================
    try:
        score_2 = 0.0
        # Check setup labels exist
        setup_labels = {
            'A3': 'Control Price',
            'A4': 'Test Price',
            'A5': 'Test Duration',
        }
        setup_values = {
            'B3': '49.99',
            'B4': '39.99',
            'B5': '30',
        }
        labels_found = 0
        for cell_ref, expected_substr in setup_labels.items():
            val = ws[cell_ref].value
            if val and expected_substr.lower() in str(val).lower():
                labels_found += 1

        values_found = 0
        for cell_ref, expected_substr in setup_values.items():
            val = ws[cell_ref].value
            if val and expected_substr in str(val):
                values_found += 1

        if labels_found >= 2 and values_found >= 2:
            score_2 += 0.05
            print(f"PASS: C2a — Setup data found ({labels_found} labels, {values_found} values)")
        else:
            print(f"FAIL: C2a — Expected setup data, found {labels_found} labels, {values_found} values")

        # Check yellow fill on input cells
        yellow_count = 0
        for cell_ref in ['B3', 'B4', 'B5']:
            try:
                fill_rgb = ws[cell_ref].fill.fgColor.rgb
                if fill_rgb and 'FFFF00' in str(fill_rgb).upper():
                    yellow_count += 1
            except Exception:
                pass

        if yellow_count >= 2:
            score_2 += 0.05
            print(f"PASS: C2b — Yellow fill on {yellow_count}/3 input cells")
        else:
            print(f"FAIL: C2b — Expected yellow fill on input cells, found {yellow_count}/3")

        total_score += score_2
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Headers row 7 with correct labels and bold (0.15)
    # =========================================================================
    try:
        score_3 = 0.0
        expected_headers = ['Day', 'Control Visitors', 'Control Purchases', 'Control Rate',
                           'Test Visitors', 'Test Purchases', 'Test Rate', 'Rate Diff', 'Cum. Significance']

        headers_found = 0
        headers_bold = 0
        for col_idx in range(1, 10):
            cell = ws.cell(row=7, column=col_idx)
            if cell.value and str(cell.value).strip().lower() in [h.lower() for h in expected_headers]:
                headers_found += 1
            if cell.font.bold:
                headers_bold += 1

        if headers_found >= 7:
            score_3 += 0.08
            print(f"PASS: C3a — {headers_found}/9 headers found in row 7")
        else:
            print(f"FAIL: C3a — Expected >=7 headers in row 7, found {headers_found}")

        if headers_bold >= 7:
            score_3 += 0.04
            print(f"PASS: C3b — {headers_bold}/9 headers are bold")
        else:
            print(f"FAIL: C3b — Expected >=7 bold headers, found {headers_bold}")

        # Check header fill (purple)
        try:
            h_fill = ws.cell(row=7, column=1).fill.fgColor.rgb
            if h_fill and '4B0082' in str(h_fill).upper():
                score_3 += 0.03
                print(f"PASS: C3c — Header purple fill: {h_fill}")
            else:
                print(f"FAIL: C3c — Expected purple header fill, found: {h_fill}")
        except Exception:
            print("FAIL: C3c — Could not read header fill")

        total_score += score_3
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: 30 rows of daily data rows 8-37 (0.15)
    # =========================================================================
    try:
        score_4 = 0.0
        # Check day numbers 1-30 in column A
        day_count = 0
        for r in range(8, 38):
            val = ws.cell(row=r, column=1).value
            if val is not None:
                try:
                    day_num = int(val)
                    if 1 <= day_num <= 30:
                        day_count += 1
                except (ValueError, TypeError):
                    pass

        if day_count >= 28:
            score_4 += 0.05
            print(f"PASS: C4a — {day_count}/30 day numbers found")
        else:
            print(f"FAIL: C4a — Expected >=28 day numbers, found {day_count}")

        # Check visitor data exists (columns B and E should have numeric data)
        visitor_count = 0
        for r in range(8, 38):
            b_val = ws.cell(row=r, column=2).value  # Control Visitors
            e_val = ws.cell(row=r, column=5).value  # Test Visitors
            if b_val is not None and e_val is not None:
                try:
                    bv = int(b_val)
                    ev = int(e_val)
                    if 100 <= bv <= 2000 and 100 <= ev <= 2000:
                        visitor_count += 1
                except (ValueError, TypeError):
                    pass

        if visitor_count >= 25:
            score_4 += 0.05
            print(f"PASS: C4b — {visitor_count}/30 rows have realistic visitor data")
        else:
            print(f"FAIL: C4b — Expected >=25 rows with visitor data, found {visitor_count}")

        # Check purchase data exists (columns C and F)
        purchase_count = 0
        for r in range(8, 38):
            c_val = ws.cell(row=r, column=3).value
            f_val = ws.cell(row=r, column=6).value
            if c_val is not None and f_val is not None:
                try:
                    cv = int(c_val)
                    fv = int(f_val)
                    if 0 <= cv <= 200 and 0 <= fv <= 200:
                        purchase_count += 1
                except (ValueError, TypeError):
                    pass

        if purchase_count >= 25:
            score_4 += 0.05
            print(f"PASS: C4c — {purchase_count}/30 rows have purchase data")
        else:
            print(f"FAIL: C4c — Expected >=25 rows with purchase data, found {purchase_count}")

        total_score += score_4
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Formulas and number formats in D, G, H columns (0.10)
    # =========================================================================
    try:
        score_5 = 0.0
        # Check D8 has formula referencing C8/B8 (control rate)
        d8_val = ws['D8'].value
        if d8_val and isinstance(d8_val, str) and '=' in d8_val and ('C8' in d8_val.upper() or 'B8' in d8_val.upper()):
            score_5 += 0.02
            print(f"PASS: C5a — D8 has rate formula: {repr(d8_val)}")
        else:
            print(f"FAIL: C5a — Expected rate formula in D8, found: {repr(d8_val)}")

        # Check G8 has formula referencing F8/E8 (test rate)
        g8_val = ws['G8'].value
        if g8_val and isinstance(g8_val, str) and '=' in g8_val and ('F8' in g8_val.upper() or 'E8' in g8_val.upper()):
            score_5 += 0.02
            print(f"PASS: C5b — G8 has rate formula: {repr(g8_val)}")
        else:
            print(f"FAIL: C5b — Expected rate formula in G8, found: {repr(g8_val)}")

        # Check H8 has rate diff formula
        h8_val = ws['H8'].value
        if h8_val and isinstance(h8_val, str) and '=' in h8_val:
            score_5 += 0.02
            print(f"PASS: C5c — H8 has diff formula: {repr(h8_val[:40])}")
        else:
            print(f"FAIL: C5c — Expected diff formula in H8, found: {repr(h8_val)}")

        # Check percentage number format on D8
        d8_nf = ws['D8'].number_format
        if d8_nf and '%' in str(d8_nf):
            score_5 += 0.02
            print(f"PASS: C5d — D8 percentage format: {repr(d8_nf)}")
        else:
            print(f"FAIL: C5d — Expected percentage format on D8, found: {repr(d8_nf)}")

        # Check H8 has +/- format
        h8_nf = ws['H8'].number_format
        if h8_nf and '%' in str(h8_nf):
            score_5 += 0.02
            print(f"PASS: C5e — H8 percentage format: {repr(h8_nf)}")
        else:
            print(f"FAIL: C5e — Expected percentage format on H8, found: {repr(h8_nf)}")

        total_score += score_5
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Conditional formatting on H and I columns (0.10)
    # =========================================================================
    try:
        score_6 = 0.0
        cf_ranges = []
        for cf in ws.conditional_formatting:
            cf_ranges.append(str(cf))

        # Check H column has conditional formatting
        h_cf_found = any('H' in r for r in cf_ranges)
        if h_cf_found:
            score_6 += 0.05
            print(f"PASS: C6a — Conditional formatting found on H column")
        else:
            print(f"FAIL: C6a — No conditional formatting on H column, ranges: {cf_ranges}")

        # Check I column has conditional formatting
        i_cf_found = any('I' in r for r in cf_ranges)
        if i_cf_found:
            score_6 += 0.05
            print(f"PASS: C6b — Conditional formatting found on I column")
        else:
            print(f"FAIL: C6b — No conditional formatting on I column, ranges: {cf_ranges}")

        total_score += score_6
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # =========================================================================
    # Component 7: Final Results row 39 + summary metrics (0.10)
    # =========================================================================
    try:
        score_7 = 0.0
        a39_val = ws['A39'].value
        if a39_val and 'Final Results' in str(a39_val):
            # Check bold and size
            if ws['A39'].font.bold:
                score_7 += 0.03
                print(f"PASS: C7a — 'Final Results' bold in A39")
            else:
                print(f"FAIL: C7a — 'Final Results' found but not bold")

            # Check double border
            border = ws['A39'].border
            if (border.top and border.top.style == 'double') or (border.bottom and border.bottom.style == 'double'):
                score_7 += 0.02
                print(f"PASS: C7b — Double border on row 39")
            else:
                print(f"FAIL: C7b — Expected double border, top={border.top.style if border.top else None}, bottom={border.bottom.style if border.bottom else None}")
        else:
            print(f"FAIL: C7a — Expected 'Final Results' in A39, found: {repr(a39_val)}")
            print(f"FAIL: C7b — No 'Final Results' row to check border")

        # Check summary metrics exist (rows 40-44)
        summary_count = 0
        for r in range(40, 45):
            for c in range(1, 10):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    summary_count += 1
                    break  # count row, not cells

        if summary_count >= 3:
            score_7 += 0.05
            print(f"PASS: C7c — {summary_count} summary metric rows found (40-44)")
        else:
            print(f"FAIL: C7c — Expected >=3 summary rows, found {summary_count}")

        total_score += score_7
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # =========================================================================
    # Component 8: Two charts - line chart and bar chart (0.15)
    # =========================================================================
    try:
        score_8 = 0.0
        charts = ws._charts

        if len(charts) >= 2:
            score_8 += 0.03
            print(f"PASS: C8a — {len(charts)} charts found (expected >=2)")
        else:
            print(f"FAIL: C8a — Expected >=2 charts, found {len(charts)}")

        # Check for line chart
        line_charts = [c for c in charts if isinstance(c, LineChart)]
        if len(line_charts) >= 1:
            score_8 += 0.04
            print(f"PASS: C8b — Line chart found")

            # Check line chart title
            try:
                title_text = ''
                for p in line_charts[0].title.tx.rich.paragraphs:
                    if hasattr(p, 'r') and p.r:
                        for run in p.r:
                            title_text += run.t
                if 'Conversion' in title_text or 'conversion' in title_text:
                    score_8 += 0.02
                    print(f"PASS: C8c — Line chart title: {repr(title_text)}")
                else:
                    print(f"FAIL: C8c — Expected 'Conversion' in line chart title, got: {repr(title_text)}")
            except Exception as e:
                print(f"FAIL: C8c — Could not read line chart title: {e}")
        else:
            print(f"FAIL: C8b — No line chart found")
            print(f"FAIL: C8c — No line chart to check title")

        # Check for bar chart
        bar_charts = [c for c in charts if isinstance(c, BarChart)]
        if len(bar_charts) >= 1:
            score_8 += 0.04
            print(f"PASS: C8d — Bar chart found")

            # Check bar chart title
            try:
                title_text = ''
                for p in bar_charts[0].title.tx.rich.paragraphs:
                    if hasattr(p, 'r') and p.r:
                        for run in p.r:
                            title_text += run.t
                if 'Cumulative' in title_text or 'cumulative' in title_text or 'Conversion' in title_text.lower():
                    score_8 += 0.02
                    print(f"PASS: C8e — Bar chart title: {repr(title_text)}")
                else:
                    print(f"FAIL: C8e — Expected 'Cumulative' in bar chart title, got: {repr(title_text)}")
            except Exception as e:
                print(f"FAIL: C8e — Could not read bar chart title: {e}")
        else:
            print(f"FAIL: C8d — No bar chart found")
            print(f"FAIL: C8e — No bar chart to check title")

        total_score += score_8
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
