"""
Initial Setup: Sales Budget vs Actual Comparison
Task ID: calc_sales_budget_actual_068
Domain: libreoffice_calc

Creates a spreadsheet with a 'BudgetActual' sheet containing:
- Headers: Category, Budget, Actual, Variance, Variance %, Status
- 11 expense categories with budget and actual values
- Columns D, E, F empty (to be filled by the agent)
- Row 13 empty (totals row to be added by the agent)
- No chart (to be created by the agent)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_budget_actual_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: BudgetActual ---
    ws = wb.active
    ws.title = 'BudgetActual'

    # Headers
    headers = ['Category', 'Budget', 'Actual', 'Variance', 'Variance %', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 11 expense categories with realistic budget and actual data
    # Budget ranges from ~$15,000 to ~$850,000
    expense_data = [
        ('Personnel',         850000,  823450),
        ('Marketing',         120000,  138760),
        ('Software',           75000,   68900),
        ('Hardware',           45000,   51230),
        ('Travel',             60000,   42180),
        ('Events',             35000,   38950),
        ('Training',           25000,   19750),
        ('Office Supplies',    15000,   16320),
        ('Consulting',         90000,   87500),
        ('Facilities',        110000,  108200),
        ('Miscellaneous',      30000,   27640),
    ]

    for r, (category, budget, actual) in enumerate(expense_data, 2):
        ws.cell(row=r, column=1, value=category)
        ws.cell(row=r, column=2, value=budget)
        ws.cell(row=r, column=3, value=actual)
        # Columns D (Variance), E (Variance %), F (Status) — EMPTY

        # Format budget and actual as currency
        ws.cell(row=r, column=2).number_format = '$#,##0'
        ws.cell(row=r, column=3).number_format = '$#,##0'

        # Alternate row shading for readability
        if r % 2 == 0:
            for col in range(1, 7):
                ws.cell(row=r, column=col).fill = PatternFill(
                    start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid'
                )

    # Row 13 is intentionally left empty (to be filled by agent as TOTAL row)

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    # Freeze row 1
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets:', wb.sheetnames)
    print('Rows with data: 1 (header) + 11 (categories) = 12 rows; row 13 empty')


create_initial()
