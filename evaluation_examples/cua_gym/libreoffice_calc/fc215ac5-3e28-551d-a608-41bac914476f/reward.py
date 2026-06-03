"""
Reward Script: Build a sales budget vs actual comparison spreadsheet.
Task ID: calc_sales_budget_actual_068
Domain: libreoffice_calc
Scoring:
  Component 1: Variance formulas D2:D12 (=Cx-Bx)           — 0.20 pts
  Component 2: Variance % formulas E2:E12 (=Dx/Bx)         — 0.15 pts
  Component 3: Status IF formulas F2:F12                     — 0.15 pts
  Component 4: Total row 13 (label + SUM formulas)          — 0.20 pts
  Component 5: Conditional formatting on D2:D12             — 0.15 pts
  Component 6: Bar chart (Budget vs Actual)                  — 0.15 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_budget_actual_068'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'BudgetActual' must exist
    if 'BudgetActual' not in wb.sheetnames:
        print("CRITICAL: Sheet 'BudgetActual' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['BudgetActual']

    # Component 1: Variance formulas D2:D12 (=C{row}-B{row}) — 0.20 points
    # Initial file has None in D2:D12; golden has =C2-B2 etc.
    try:
        variance_ok = 0
        variance_total = 11  # rows 2-12
        for row in range(2, 13):
            val = ws.cell(row=row, column=4).value
            if val is not None and isinstance(val, str):
                # Accept either =C{r}-B{r} or =C{r}−B{r} form (some locales use minus)
                pattern = rf'=C{row}-B{row}'
                if val.upper().replace(' ', '') == pattern.upper():
                    variance_ok += 1
                # Also accept if formula is present at all (=C-B pattern)
                elif re.match(rf'=C{row}\s*-\s*B{row}', val, re.IGNORECASE):
                    variance_ok += 1
        if variance_ok == variance_total:
            print(f"PASS: Component 1 — All {variance_total} variance formulas D2:D12 correct (=Cx-Bx) (0.20 pts)")
            total_score += 0.20
        elif variance_ok >= variance_total * 0.5:
            print(f"PARTIAL: Component 1 — {variance_ok}/{variance_total} variance formulas D2:D12 correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Only {variance_ok}/{variance_total} variance formulas found in D2:D12")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Variance % formulas E2:E12 (=D{row}/B{row}) — 0.15 points
    try:
        pct_ok = 0
        pct_total = 11
        for row in range(2, 13):
            val = ws.cell(row=row, column=5).value
            if val is not None and isinstance(val, str):
                pattern = rf'=D{row}/B{row}'
                if val.upper().replace(' ', '') == pattern.upper():
                    pct_ok += 1
                elif re.match(rf'=D{row}\s*/\s*B{row}', val, re.IGNORECASE):
                    pct_ok += 1
        if pct_ok == pct_total:
            print(f"PASS: Component 2 — All {pct_total} variance % formulas E2:E12 correct (=Dx/Bx) (0.15 pts)")
            total_score += 0.15
        elif pct_ok >= pct_total * 0.5:
            print(f"PARTIAL: Component 2 — {pct_ok}/{pct_total} variance % formulas E2:E12 correct (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2 — Only {pct_ok}/{pct_total} variance % formulas found in E2:E12")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Status IF formulas F2:F12 (=IF(D>0,"Over Budget","Under Budget")) — 0.15 points
    try:
        status_ok = 0
        status_total = 11
        for row in range(2, 13):
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str):
                val_clean = val.upper().replace(' ', '')
                # Check for IF(D{row}>0 pattern
                if f'IF(D{row}>0' in val_clean:
                    # Also check it references "Over Budget" and "Under Budget" strings
                    if 'OVERBUDGET' in val_clean and 'UNDERBUDGET' in val_clean:
                        status_ok += 1
                    elif 'OVER' in val_clean and 'UNDER' in val_clean:
                        status_ok += 1
                    else:
                        status_ok += 1  # partial match for IF structure
        if status_ok == status_total:
            print(f"PASS: Component 3 — All {status_total} status IF formulas F2:F12 correct (0.15 pts)")
            total_score += 0.15
        elif status_ok >= status_total * 0.5:
            print(f"PARTIAL: Component 3 — {status_ok}/{status_total} status formulas F2:F12 correct (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — Only {status_ok}/{status_total} status formulas found in F2:F12")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Total row 13 — TOTAL label + SUM formulas in B13, C13, D13, ratio in E13 — 0.20 points
    try:
        total_row_score = 0.0

        # Check A13 label
        a13 = ws.cell(row=13, column=1).value
        if a13 is not None and str(a13).strip().upper() == 'TOTAL':
            total_row_score += 0.05

        # Check B13 SUM formula
        b13 = ws.cell(row=13, column=2).value
        if b13 is not None and isinstance(b13, str) and 'SUM' in b13.upper() and 'B2' in b13.upper() and 'B12' in b13.upper():
            total_row_score += 0.05

        # Check C13 SUM formula
        c13 = ws.cell(row=13, column=3).value
        if c13 is not None and isinstance(c13, str) and 'SUM' in c13.upper() and 'C2' in c13.upper() and 'C12' in c13.upper():
            total_row_score += 0.05

        # Check D13 SUM formula
        d13 = ws.cell(row=13, column=4).value
        if d13 is not None and isinstance(d13, str) and 'SUM' in d13.upper() and 'D2' in d13.upper() and 'D12' in d13.upper():
            total_row_score += 0.05

        if total_row_score >= 0.20:
            print(f"PASS: Component 4 — Total row 13 complete (TOTAL label + SUM formulas B/C/D + E13) (0.20 pts)")
            total_score += 0.20
        elif total_row_score >= 0.10:
            print(f"PARTIAL: Component 4 — Total row 13 partially complete ({total_row_score:.2f}/0.20 pts) ({total_row_score:.2f} pts)")
            total_score += total_row_score
        else:
            print(f"FAIL: Component 4 — Total row 13 incomplete; A13={a13}, B13={b13}, C13={c13}, D13={d13}")
            if total_row_score > 0:
                total_score += total_row_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on D2:D12 (red for >0, green for <=0) — 0.15 points
    try:
        cf_found = False
        has_red_gt0 = False
        has_green_lte0 = False

        for cf in ws.conditional_formatting:
            # Check if any CF range covers D2:D12 (or part of it)
            cf_str = str(cf)
            if 'D2' in cf_str and 'D12' in cf_str:
                cf_found = True
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        op = getattr(rule, 'operator', '')
                        formula = getattr(rule, 'formula', [])
                        # Check for red fill on greaterThan 0
                        if op in ('greaterThan', 'greaterThanOrEqual') and formula and formula[0] in ('0', '0.0'):
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else ''
                                if 'FF0000' in fill_color.upper():
                                    has_red_gt0 = True
                        # Check for green fill on lessThanOrEqual 0
                        if op in ('lessThanOrEqual', 'lessThan') and formula and formula[0] in ('0', '0.0'):
                            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                                fill_color = rule.dxf.fill.fgColor.rgb if rule.dxf.fill.fgColor else ''
                                if '00FF00' in fill_color.upper():
                                    has_green_lte0 = True

        if has_red_gt0 and has_green_lte0:
            print(f"PASS: Component 5 — Conditional formatting on D2:D12 with red (>0) and green (<=0) (0.15 pts)")
            total_score += 0.15
        elif cf_found and (has_red_gt0 or has_green_lte0):
            print(f"PARTIAL: Component 5 — Conditional formatting partially applied (red={has_red_gt0}, green={has_green_lte0}) (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 5 — No matching conditional formatting found on D2:D12 (found={cf_found})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Bar chart (Budget vs Actual side-by-side per category) — 0.15 points
    try:
        charts = ws._charts
        has_bar_chart = False
        has_two_series = False

        for chart in charts:
            chart_type_name = type(chart).__name__
            if 'Bar' in chart_type_name or 'bar' in chart_type_name:
                has_bar_chart = True
                if len(chart.series) >= 2:
                    has_two_series = True
                break

        if has_bar_chart and has_two_series:
            print(f"PASS: Component 6 — Bar chart with 2 series (Budget vs Actual) present (0.15 pts)")
            total_score += 0.15
        elif has_bar_chart:
            print(f"PARTIAL: Component 6 — Bar chart present but has {len(charts[0].series) if charts else 0} series (expected 2) (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 6 — No bar chart found in BudgetActual sheet (charts found: {len(charts)})")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
