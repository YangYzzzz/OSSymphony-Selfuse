"""
Initial Setup: Remove data validation from cell D5 only
Task ID: calc_nrv_071
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    headers = ['Product ID', 'Product Name', 'Category', 'Quantity']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Data rows (realistic inventory data)
    data = [
        ['INV-1001', 'Wireless Bluetooth Headphones', 'Electronics', 245],
        ['INV-1002', 'Organic Green Tea (100 bags)', 'Beverages', 512],
        ['INV-1003', 'Stainless Steel Water Bottle', 'Kitchen', 178],
        ['INV-1004', 'Ergonomic Office Chair', 'Furniture', 63],
        ['INV-1005', 'USB-C Charging Cable (6ft)', 'Electronics', 890],
        ['INV-1006', 'Bamboo Cutting Board Set', 'Kitchen', 324],
        ['INV-1007', 'LED Desk Lamp with Dimmer', 'Lighting', 156],
        ['INV-1008', 'Premium Notebook (A5, Lined)', 'Stationery', 743],
        ['INV-1009', 'Adjustable Monitor Stand', 'Furniture', 92],
        ['INV-1010', 'Cold Brew Coffee Maker', 'Kitchen', 201],
        ['INV-1011', 'Mechanical Keyboard (TKL)', 'Electronics', 417],
        ['INV-1012', 'Yoga Mat (6mm Thick)', 'Fitness', 289],
        ['INV-1013', 'Portable Power Bank 20000mAh', 'Electronics', 634],
        ['INV-1014', 'Cast Iron Skillet (12 inch)', 'Kitchen', 147],
        ['INV-1015', 'Noise Cancelling Earbuds', 'Electronics', 523],
        ['INV-1016', 'Standing Desk Converter', 'Furniture', 38],
        ['INV-1017', 'Reusable Grocery Bags (Set of 5)', 'Household', 956],
        ['INV-1018', 'Ceramic Coffee Mug (16oz)', 'Kitchen', 412],
        ['INV-1019', 'Wireless Mouse (Ergonomic)', 'Electronics', 367],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12

    # Add whole number data validation to D2:D20 (1-1000)
    dv = DataValidation(
        type='whole',
        operator='between',
        formula1='1',
        formula2='1000',
        allow_blank=False,
    )
    dv.error = 'Please enter a whole number between 1 and 1000.'
    dv.errorTitle = 'Invalid Quantity'
    dv.prompt = 'Enter quantity (1-1000)'
    dv.promptTitle = 'Quantity'
    dv.showErrorMessage = True
    dv.showInputMessage = True
    dv.add('D2:D20')
    ws.add_data_validation(dv)

    # --- Sheet 2: Categories ---
    ws2 = wb.create_sheet('Categories')
    cat_headers = ['Category', 'Department', 'Manager']
    for col, h in enumerate(cat_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    cat_data = [
        ['Electronics', 'Technology', 'James Rivera'],
        ['Beverages', 'Food & Drink', 'Priya Sharma'],
        ['Kitchen', 'Home Goods', 'Mei-Ling Wu'],
        ['Furniture', 'Office', 'David Okonkwo'],
        ['Lighting', 'Home Goods', 'Mei-Ling Wu'],
        ['Stationery', 'Office', 'David Okonkwo'],
        ['Fitness', 'Sports', 'Aisha Patel'],
        ['Household', 'Home Goods', 'Mei-Ling Wu'],
    ]
    for r, row_data in enumerate(cat_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
