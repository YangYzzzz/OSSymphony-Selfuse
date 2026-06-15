"""
Reward Script: XY Scatter Chart with 4 Customer Segment Series
Task ID: calc_gcp_089
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - A scatter chart exists on the CustomerSegments sheet
  Component 2 (0.20) - Chart has exactly 4 data series
  Component 3 (0.20) - Series titles match the 4 expected segments
  Component 4 (0.20) - Each series uses a distinct marker shape
  Component 5 (0.20) - Each series uses a distinct color
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_089'
EXPECTED_SEGMENTS = {'Premium', 'Regular', 'Occasional', 'New'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the CustomerSegments sheet
    target_sheet = None
    for name in wb.sheetnames:
        if name.lower().replace(' ', '') == 'customersegments':
            target_sheet = wb[name]
            break
    if target_sheet is None:
        target_sheet = wb.active

    ws = target_sheet

    # Precondition: sheet must have data (at least header + some rows)
    if ws.max_row < 5:
        print("PRECONDITION FAIL: Sheet has too few rows — not the expected data sheet")
        print("REWARD: 0.0")
        return 0.0

    # ----------------------------------------------------------------
    # Component 1: A scatter chart exists (0.20 points)
    # Initial has 0 charts, golden should have >= 1 scatter chart
    # ----------------------------------------------------------------
    scatter_chart = None
    try:
        charts = ws._charts
        # Look for ScatterChart specifically
        from openpyxl.chart import ScatterChart
        scatter_charts = [c for c in charts if isinstance(c, ScatterChart)]
        if len(scatter_charts) >= 1:
            scatter_chart = scatter_charts[0]
            print(f"PASS: Component 1 — Scatter chart found ({len(scatter_charts)} scatter chart(s)) (0.20 pts)")
            total_score += 0.20
        else:
            # Also check other sheets for a scatter chart
            for sn in wb.sheetnames:
                other_ws = wb[sn]
                other_scatter = [c for c in other_ws._charts if isinstance(c, ScatterChart)]
                if other_scatter:
                    scatter_chart = other_scatter[0]
                    print(f"PASS: Component 1 — Scatter chart found on sheet '{sn}' (0.20 pts)")
                    total_score += 0.20
                    break
            else:
                print(f"FAIL: Component 1 — No scatter chart found. Total charts across all sheets: {sum(len(wb[s]._charts) for s in wb.sheetnames)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # If no scatter chart found, remaining checks cannot proceed
    if scatter_chart is None:
        print("No scatter chart to evaluate further components.")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # ----------------------------------------------------------------
    # Component 2: Chart has exactly 4 data series (0.20 points)
    # ----------------------------------------------------------------
    try:
        num_series = len(scatter_chart.series)
        if num_series == 4:
            print(f"PASS: Component 2 — Chart has 4 data series (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Expected 4 series, found {num_series}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Series titles match the 4 expected segments (0.20 points)
    # ----------------------------------------------------------------
    try:
        series_titles = set()
        for s in scatter_chart.series:
            title_val = None
            if s.title:
                # title can be a SeriesLabel with .v attribute or .strRef
                if hasattr(s.title, 'v') and s.title.v:
                    title_val = s.title.v
                elif hasattr(s.title, 'value') and s.title.value:
                    title_val = s.title.value
                elif isinstance(s.title, str):
                    title_val = s.title
            if title_val:
                series_titles.add(title_val.strip())

        # Check if all 4 expected segments are represented
        matched = EXPECTED_SEGMENTS.intersection(series_titles)
        if len(matched) == 4:
            print(f"PASS: Component 3 — All 4 segment titles found: {sorted(matched)} (0.20 pts)")
            total_score += 0.20
        elif len(matched) >= 2:
            partial = 0.20 * (len(matched) / 4)
            print(f"PARTIAL: Component 3 — {len(matched)}/4 segments matched: {sorted(matched)} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Expected segments {EXPECTED_SEGMENTS}, found titles: {series_titles}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ----------------------------------------------------------------
    # Component 4: Each series uses a distinct marker shape (0.20 points)
    # ----------------------------------------------------------------
    try:
        marker_symbols = []
        for s in scatter_chart.series:
            symbol = None
            if hasattr(s, 'marker') and s.marker and s.marker.symbol:
                symbol = s.marker.symbol
            marker_symbols.append(symbol)

        # Filter out None values
        valid_symbols = [m for m in marker_symbols if m is not None]
        unique_symbols = set(valid_symbols)

        if len(valid_symbols) == 4 and len(unique_symbols) == 4:
            print(f"PASS: Component 4 — 4 distinct marker shapes: {sorted(unique_symbols)} (0.20 pts)")
            total_score += 0.20
        elif len(valid_symbols) >= 2 and len(unique_symbols) >= 2:
            # Partial: at least some distinct markers
            ratio = len(unique_symbols) / 4
            partial = 0.20 * ratio
            print(f"PARTIAL: Component 4 — {len(unique_symbols)} distinct markers out of 4: {sorted(unique_symbols)} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Expected 4 distinct markers, found symbols: {marker_symbols}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ----------------------------------------------------------------
    # Component 5: Each series uses a distinct color (0.20 points)
    # ----------------------------------------------------------------
    try:
        series_colors = []
        for s in scatter_chart.series:
            color = None
            # Check marker fill color first (common for scatter charts)
            if hasattr(s, 'marker') and s.marker:
                mgp = s.marker.graphicalProperties
                if mgp and mgp.solidFill:
                    sf = mgp.solidFill
                    if hasattr(sf, 'srgbClr') and sf.srgbClr:
                        color = str(sf.srgbClr)
            # Fallback to series-level graphicalProperties solidFill
            if color is None and s.graphicalProperties and s.graphicalProperties.solidFill:
                sf = s.graphicalProperties.solidFill
                if hasattr(sf, 'srgbClr') and sf.srgbClr:
                    color = str(sf.srgbClr)
            series_colors.append(color)

        valid_colors = [c for c in series_colors if c is not None]
        unique_colors = set(valid_colors)

        if len(valid_colors) == 4 and len(unique_colors) == 4:
            print(f"PASS: Component 5 — 4 distinct colors: {sorted(unique_colors)} (0.20 pts)")
            total_score += 0.20
        elif len(valid_colors) >= 2 and len(unique_colors) >= 2:
            ratio = len(unique_colors) / 4
            partial = 0.20 * ratio
            print(f"PARTIAL: Component 5 — {len(unique_colors)} distinct colors out of 4: {sorted(unique_colors)} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Expected 4 distinct colors, found: {series_colors}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
