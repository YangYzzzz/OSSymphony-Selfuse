"""
Initial Setup: Create CustomerSegments spreadsheet with 120 rows of customer data
Task ID: calc_gcp_089
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_089'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CustomerSegments'

    # Headers
    headers = ['CustomerID', 'Segment', 'PurchaseFrequency', 'AvgOrderValue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Define segment patterns
    segments = {
        'Premium': {'freq_range': (25, 52), 'value_range': (150, 350), 'count': 30},
        'Regular': {'freq_range': (12, 30), 'value_range': (60, 150), 'count': 30},
        'Occasional': {'freq_range': (3, 12), 'value_range': (15, 80), 'count': 30},
        'New': {'freq_range': (1, 8), 'value_range': (15, 200), 'count': 30},
    }

    row_idx = 2
    cust_num = 1
    for seg_name, params in segments.items():
        for _ in range(params['count']):
            cid = f'C{cust_num:03d}'
            freq = random.randint(params['freq_range'][0], params['freq_range'][1])
            val = round(random.uniform(params['value_range'][0], params['value_range'][1]), 2)
            ws.cell(row=row_idx, column=1, value=cid)
            ws.cell(row=row_idx, column=2, value=seg_name)
            ws.cell(row=row_idx, column=3, value=freq)
            ws.cell(row=row_idx, column=4, value=val)
            row_idx += 1
            cust_num += 1

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
