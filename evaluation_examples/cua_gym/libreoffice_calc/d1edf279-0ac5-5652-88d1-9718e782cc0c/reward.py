"""
Reward Script: Update pivot table source range from A1:F101 to A1:F251 and refresh
Task ID: calc_pivot_049
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Grand total updated to 235000
  Component 2 (0.4): Individual category sums updated to reflect full 250 rows
  Component 3 (0.2): All 5 categories present with correct updated values
"""

import os
import time


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_049'

# Expected golden values for pivot table after refresh with full 250 rows
EXPECTED_CATEGORIES = {
    'Clothing': 45500,
    'Electronics': 70500,
    'Food': 30000,
    'Furniture': 55000,
    'Stationery': 34000,
}
EXPECTED_GRAND_TOTAL = 235000

# Initial values (before task - pivot only covered rows 1:101)
INITIAL_GRAND_TOTAL = 95000


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Report sheet must exist
    if 'Report' not in wb.sheetnames:
        print("FAIL: 'Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Build a map of category -> value from the Report sheet
    # The pivot table has Category in column A, Sum of Revenue in column B
    # Starting from row 4 (row 3 is header: Category, Sum - Revenue)
    category_values = {}
    grand_total_value = None

    try:
        for r in range(4, ws.max_row + 1):
            cat = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if cat is not None and val is not None:
                cat_str = str(cat).strip()
                if cat_str.lower() in ('total result', 'grand total', 'total'):
                    grand_total_value = val
                else:
                    category_values[cat_str] = val
        print(f"INFO: Found categories: {category_values}")
        print(f"INFO: Grand total: {grand_total_value}")
    except Exception as e:
        print(f"ERROR: Could not parse Report sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Grand total updated to 235000 (0.4 points)
    # This MUST fail on initial (95000) and pass on golden (235000)
    try:
        if grand_total_value is not None:
            gt_num = float(grand_total_value)
            # Check that grand total is close to 235000 (with tolerance for rounding)
            if abs(gt_num - EXPECTED_GRAND_TOTAL) < 500:
                print(f"PASS: Component 1 - Grand total is {gt_num}, expected ~{EXPECTED_GRAND_TOTAL} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 - Grand total is {gt_num}, expected ~{EXPECTED_GRAND_TOTAL}")
        else:
            print("FAIL: Component 1 - No grand total row found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: At least 3 of 5 categories have updated values significantly higher than initial (0.4 points)
    # Initial values were: Clothing=18500, Electronics=28500, Food=12000, Furniture=22000, Stationery=14000
    # Golden values are approximately 2.5x the initial values
    # This MUST fail on initial and pass on golden
    try:
        updated_count = 0
        for cat, expected_val in EXPECTED_CATEGORIES.items():
            if cat in category_values:
                actual_val = float(category_values[cat])
                # Check that value is close to the expected golden value (within 10% tolerance)
                if abs(actual_val - expected_val) / expected_val < 0.10:
                    updated_count += 1
                    print(f"  MATCH: {cat} = {actual_val} (expected ~{expected_val})")
                else:
                    print(f"  MISMATCH: {cat} = {actual_val} (expected ~{expected_val})")
            else:
                print(f"  MISSING: {cat} not found in pivot table")

        if updated_count >= 4:
            print(f"PASS: Component 2 - {updated_count}/5 categories have updated values (0.4 pts)")
            total_score += 0.4
        elif updated_count >= 3:
            print(f"PARTIAL: Component 2 - {updated_count}/5 categories updated (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 - Only {updated_count}/5 categories have updated values")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: All 5 categories present in the pivot table (0.2 points)
    # AND their values are significantly different from the initial state
    # This MUST fail on initial and pass on golden
    try:
        all_present = all(cat in category_values for cat in EXPECTED_CATEGORIES)
        # Check that the sum of all categories is significantly above the initial total
        if all_present:
            actual_sum = sum(float(category_values[cat]) for cat in EXPECTED_CATEGORIES)
            # Initial sum was 95000, golden sum should be ~235000
            # Only pass if sum is clearly above the initial (e.g., > 150000)
            if actual_sum > 150000:
                print(f"PASS: Component 3 - All 5 categories present, sum={actual_sum} > 150000 (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 - All categories present but sum={actual_sum} not above threshold 150000")
        else:
            missing = [c for c in EXPECTED_CATEGORIES if c not in category_values]
            print(f"FAIL: Component 3 - Missing categories: {missing}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
