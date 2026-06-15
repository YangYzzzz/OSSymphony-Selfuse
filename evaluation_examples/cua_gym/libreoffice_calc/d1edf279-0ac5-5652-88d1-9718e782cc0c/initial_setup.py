"""
Initial Setup: Pivot table with source data range A1:F101 (100 data rows)
Task ID: calc_pivot_049
Domain: libreoffice_calc

Creates a spreadsheet with 250 rows of data on 'GrowingData' sheet,
then creates a pivot table on 'Report' sheet using only the first 100 rows (A1:F101).
The pivot shows Category with SUM of Revenue, grand total = 95000.
"""

import os
import random
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_data_file():
    """Create the xlsx with 250 data rows and blank Report sheet."""
    import openpyxl

    random.seed(42)

    categories = ['Electronics', 'Furniture', 'Clothing', 'Food', 'Stationery']
    products = {
        'Electronics': ['Laptop', 'Headphones', 'Monitor', 'Keyboard', 'Mouse'],
        'Furniture': ['Desk', 'Chair', 'Shelf', 'Cabinet', 'Table'],
        'Clothing': ['Jacket', 'Shirt', 'Pants', 'Shoes', 'Hat'],
        'Food': ['Coffee', 'Snacks', 'Lunch Box', 'Water', 'Fruit'],
        'Stationery': ['Notebook', 'Pen Set', 'Markers', 'Binder', 'Tape'],
    }

    # Revenue distribution:
    #   First 100 rows sum = 95000
    #   All 250 rows sum = 235000
    #   Rows 101-250 (150 rows) sum = 140000

    # Assign revenues per category to make it realistic
    # 5 categories, each appears every 5 rows (round-robin)
    # First 100 rows: 20 rows per category, category sums add to 95000
    # All 250 rows: 50 rows per category, sums add to 235000

    # Target per-category sums for first 100 rows (total 95000):
    cat_sums_100 = {
        'Electronics': 28500,
        'Furniture': 22000,
        'Clothing': 18500,
        'Food': 12000,
        'Stationery': 14000,
    }
    # => total = 95000

    # Target per-category sums for rows 101-250 (total 140000):
    cat_sums_extra = {
        'Electronics': 42000,
        'Furniture': 33000,
        'Clothing': 27000,
        'Food': 18000,
        'Stationery': 20000,
    }
    # => total = 140000

    months_2024 = [f'2024-{m:02d}' for m in range(1, 13)]
    months_2025 = [f'2025-{m:02d}' for m in range(1, 7)]
    all_months = months_2024 + months_2025

    def generate_revenues(count, target_sum):
        """Generate 'count' random revenues summing to target_sum."""
        vals = []
        remaining = target_sum
        for i in range(count):
            left = count - i
            if left == 1:
                vals.append(round(remaining, 2))
            else:
                avg = remaining / left
                lo = max(100, avg * 0.4)
                hi = min(avg * 1.8, remaining - (left - 1) * 100)
                v = round(random.uniform(lo, hi), 2)
                vals.append(v)
                remaining -= v
        return vals

    # Build per-category revenue lists
    cat_revs_100 = {}
    cat_revs_extra = {}
    for cat in categories:
        cat_revs_100[cat] = generate_revenues(20, cat_sums_100[cat])
        cat_revs_extra[cat] = generate_revenues(30, cat_sums_extra[cat])

    # Counters per category
    cat_idx_100 = {c: 0 for c in categories}
    cat_idx_extra = {c: 0 for c in categories}

    data_rows = []
    for i in range(250):
        cat = categories[i % 5]
        prod = random.choice(products[cat])
        month = all_months[i % len(all_months)]
        day = random.randint(1, 28)
        date_str = f'{month}-{day:02d}'
        qty = random.randint(1, 50)

        if i < 100:
            revenue = cat_revs_100[cat][cat_idx_100[cat]]
            cat_idx_100[cat] += 1
        else:
            revenue = cat_revs_extra[cat][cat_idx_extra[cat]]
            cat_idx_extra[cat] += 1

        data_rows.append([i + 1, date_str, cat, prod, qty, revenue])

    # Verify sums
    sum_100 = sum(row[5] for row in data_rows[:100])
    sum_250 = sum(row[5] for row in data_rows)
    print(f'First 100 rows revenue: {sum_100:.2f} (target: 95000)')
    print(f'All 250 rows revenue: {sum_250:.2f} (target: 235000)')

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'GrowingData'

    headers = ['ID', 'Date', 'Category', 'Product', 'Qty', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    for r, row_data in enumerate(data_rows, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 14

    # Create blank Report sheet (pivot table added later via LibreOffice)
    wb.create_sheet('Report')

    wb.save(OUTPUT)
    print(f'Data file created: {OUTPUT}')
    return data_rows


def create_pivot_via_libreoffice():
    """Use LibreOffice UNO Python bridge to create pivot table."""

    # Write the UNO bridge script
    uno_script = '''
import subprocess
import time
import os
import sys

# Kill any existing LibreOffice
subprocess.run(["pkill", "-9", "-f", "soffice"], capture_output=True)
time.sleep(2)

env = os.environ.copy()
env["DISPLAY"] = ":0"

# Start LibreOffice with socket listener
proc = subprocess.Popen(
    ["soffice",
     "--headless",
     "--norestore",
     "--accept=socket,host=localhost,port=2002;urp;StarOffice.ServiceManager",
     "/home/user/calc_pivot_049.xlsx"],
    env=env,
    stdout=subprocess.DEVNULL,
    stderr=subprocess.DEVNULL,
)
time.sleep(8)

# Connect via UNO
import uno

localContext = uno.getComponentContext()
resolver = localContext.ServiceManager.createInstanceWithContext(
    "com.sun.star.bridge.UnoUrlResolver", localContext)

ctx = None
for attempt in range(15):
    try:
        ctx = resolver.resolve(
            "uno:socket,host=localhost,port=2002;urp;StarOffice.ComponentContext")
        print(f"Connected on attempt {attempt+1}")
        break
    except Exception as e:
        print(f"Attempt {attempt+1}: {e}")
        time.sleep(2)

if ctx is None:
    print("ERROR: Could not connect to LibreOffice")
    sys.exit(1)

smgr = ctx.ServiceManager
desktop = smgr.createInstanceWithContext("com.sun.star.frame.Desktop", ctx)

doc = None
for attempt in range(15):
    doc = desktop.getCurrentComponent()
    if doc is not None and hasattr(doc, 'Sheets'):
        break
    time.sleep(1)

if doc is None:
    print("ERROR: No document loaded")
    sys.exit(1)

print(f"Document loaded with {doc.Sheets.getCount()} sheets")

# Get source sheet and range
source_sheet = doc.Sheets.getByName("GrowingData")
source_range = source_sheet.getCellRangeByName("A1:F101")

# Get Report sheet index
report_idx = -1
for i in range(doc.Sheets.getCount()):
    if doc.Sheets.getByIndex(i).getName() == "Report":
        report_idx = i
        break

target_sheet = doc.Sheets.getByName("Report")

# Create DataPilot (pivot table)
dp_tables = target_sheet.getDataPilotTables()
dp_desc = dp_tables.createDataPilotDescriptor()
dp_desc.setSourceRange(source_range.getRangeAddress())

fields = dp_desc.getDataPilotFields()
for i in range(fields.getCount()):
    field = fields.getByIndex(i)
    name = field.getName()
    if name == "Category":
        field.Orientation = uno.Enum(
            "com.sun.star.sheet.DataPilotFieldOrientation", "ROW")
    elif name == "Revenue":
        field.Orientation = uno.Enum(
            "com.sun.star.sheet.DataPilotFieldOrientation", "DATA")
        field.Function = uno.Enum(
            "com.sun.star.sheet.GeneralFunction", "SUM")

from com.sun.star.table import CellAddress
output_addr = CellAddress()
output_addr.Sheet = report_idx
output_addr.Column = 0
output_addr.Row = 0

dp_tables.insertNewByName("PivotReport", output_addr, dp_desc)
print("Pivot table created")

# Read back the grand total to verify
# Pivot layout: Row headers in col A, values in col B
# Find the Total row
report = doc.Sheets.getByName("Report")
for row in range(20):
    cell_a = report.getCellByPosition(0, row)
    cell_b = report.getCellByPosition(1, row)
    print(f"  Report row {row}: A={cell_a.getString()}, B={cell_b.getValue()}")

# Save
doc.store()
print("Document saved")

# Close
doc.dispose()
time.sleep(1)
subprocess.run(["pkill", "-9", "-f", "soffice"], capture_output=True)
print("Done")
'''

    script_path = '/tmp/create_pivot.py'
    with open(script_path, 'w') as f:
        f.write(uno_script)

    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        ['python3', script_path],
        env=env,
        capture_output=True,
        text=True,
        timeout=180
    )
    print(f'UNO stdout:\n{result.stdout}')
    if result.stderr:
        print(f'UNO stderr:\n{result.stderr}')
    print(f'UNO return code: {result.returncode}')

    if result.returncode != 0:
        raise RuntimeError(f'Pivot table creation failed: {result.stderr}')


def main():
    create_data_file()
    create_pivot_via_libreoffice()

    # Kill headless LibreOffice, then open GUI version
    subprocess.run(['pkill', '-9', '-f', 'soffice'], capture_output=True)
    time.sleep(2)

    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


main()
