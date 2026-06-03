"""
Initial Setup: HR Review Schedule - NETWORKDAYS and EDATE task
Task ID: calc_hr_networkdays_workdays_013
Domain: libreoffice_calc

Creates a spreadsheet with new hire data. Columns D (Review Due Date) and E
(Working Days to Review) are left EMPTY for the agent to fill in.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_networkdays_workdays_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Review Schedule'

    # --- Headers ---
    headers = ['Emp ID', 'Name', 'Hire Date', 'Review Due Date', 'Working Days to Review', 'Department']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # --- Realistic employee data ---
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Human Resources',
        'Finance', 'Operations', 'Customer Success', 'Product Management',
        'Legal', 'Design'
    ]

    employees = [
        ('EMP-1001', 'Sarah Chen',          date(2024, 1, 8),  'Engineering'),
        ('EMP-1002', 'Marcus Johnson',       date(2024, 1, 15), 'Marketing'),
        ('EMP-1003', 'Priya Patel',          date(2024, 1, 22), 'Sales'),
        ('EMP-1004', 'Liam O\'Brien',        date(2024, 2, 5),  'Human Resources'),
        ('EMP-1005', 'Aisha Williams',       date(2024, 2, 12), 'Finance'),
        ('EMP-1006', 'Diego Ramirez',        date(2024, 2, 19), 'Engineering'),
        ('EMP-1007', 'Mei-Ling Zhang',       date(2024, 2, 26), 'Product Management'),
        ('EMP-1008', 'James Okafor',         date(2024, 3, 4),  'Operations'),
        ('EMP-1009', 'Natasha Petrov',       date(2024, 3, 11), 'Customer Success'),
        ('EMP-1010', 'Carlos Mendez',        date(2024, 3, 18), 'Legal'),
        ('EMP-1011', 'Emily Larsen',         date(2024, 3, 25), 'Design'),
        ('EMP-1012', 'Kwame Asante',         date(2024, 4, 1),  'Engineering'),
        ('EMP-1013', 'Sophie Moreau',        date(2024, 4, 8),  'Sales'),
        ('EMP-1014', 'Ravi Krishnamurthy',   date(2024, 4, 15), 'Finance'),
        ('EMP-1015', 'Ingrid Hoffmann',      date(2024, 4, 22), 'Marketing'),
        ('EMP-1016', 'Tobias Müller',        date(2024, 4, 29), 'Engineering'),
        ('EMP-1017', 'Fatima Al-Hassan',     date(2024, 5, 6),  'Human Resources'),
        ('EMP-1018', 'Brandon Lee',          date(2024, 5, 13), 'Product Management'),
        ('EMP-1019', 'Amara Diallo',         date(2024, 5, 20), 'Customer Success'),
        ('EMP-1020', 'Isabella Rossi',       date(2024, 5, 27), 'Operations'),
        ('EMP-1021', 'Hiro Tanaka',          date(2024, 6, 3),  'Engineering'),
        ('EMP-1022', 'Valentina Cruz',       date(2024, 6, 10), 'Sales'),
        ('EMP-1023', 'Ethan Kowalski',       date(2024, 6, 17), 'Finance'),
        ('EMP-1024', 'Nia Osei',             date(2024, 6, 24), 'Design'),
        ('EMP-1025', 'Aleksei Volkov',       date(2024, 7, 1),  'Engineering'),
        ('EMP-1026', 'Camille Dubois',       date(2024, 7, 8),  'Marketing'),
        ('EMP-1027', 'Tariq Abdullah',       date(2024, 7, 15), 'Legal'),
        ('EMP-1028', 'Yuki Nakamura',        date(2024, 7, 22), 'Product Management'),
        ('EMP-1029', 'Zoe Thompson',         date(2024, 7, 29), 'Customer Success'),
        ('EMP-1030', 'Samuel Abebe',         date(2024, 8, 5),  'Operations'),
        ('EMP-1031', 'Elena Vasquez',        date(2024, 8, 12), 'Engineering'),
        ('EMP-1032', 'Patrick Fitzgerald',   date(2024, 8, 19), 'Sales'),
        ('EMP-1033', 'Mei Huang',            date(2024, 8, 26), 'Finance'),
        ('EMP-1034', 'Olumide Adeleke',      date(2024, 9, 2),  'Human Resources'),
        ('EMP-1035', 'Larissa Santos',       date(2024, 9, 9),  'Marketing'),
        ('EMP-1036', 'Viktor Sorokin',       date(2024, 9, 16), 'Engineering'),
        ('EMP-1037', 'Ananya Sharma',        date(2024, 9, 23), 'Design'),
        ('EMP-1038', 'Michael Okonkwo',      date(2024, 9, 30), 'Product Management'),
        ('EMP-1039', 'Chiara Bianchi',       date(2024, 10, 7), 'Customer Success'),
        ('EMP-1040', 'Jacob Andersen',       date(2024, 10, 14), 'Finance'),
        ('EMP-1041', 'Nadia Karim',          date(2024, 10, 21), 'Legal'),
        ('EMP-1042', 'Seun Adeyemi',         date(2024, 10, 28), 'Operations'),
        ('EMP-1043', 'Lily Bergström',       date(2024, 11, 4), 'Engineering'),
    ]

    for row_idx, (emp_id, name, hire_date, dept) in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=name)
        # Store hire date as a date value with date format
        hire_cell = ws.cell(row=row_idx, column=3, value=hire_date)
        hire_cell.number_format = 'yyyy-mm-dd'
        # Columns D and E left EMPTY (task requirement)
        ws.cell(row=row_idx, column=4, value=None)
        ws.cell(row=row_idx, column=5, value=None)
        ws.cell(row=row_idx, column=6, value=dept)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 20

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Review Schedule')
    print(f'  Rows: 1 header + 43 data rows (2-44)')
    print(f'  Columns D and E are empty (to be filled by agent)')


create_initial()
