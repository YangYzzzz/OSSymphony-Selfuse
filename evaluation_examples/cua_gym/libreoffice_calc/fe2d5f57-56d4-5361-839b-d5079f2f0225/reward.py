"""
Reward Script: Calculate working days between hire date and review date using NETWORKDAYS and EDATE
Task ID: calc_hr_networkdays_workdays_013
Domain: libreoffice_calc
Scoring:
  - Component 1: D2:D44 contains =EDATE(Cx,6) formula (0.40 pts)
  - Component 2: E2:E44 contains =NETWORKDAYS(Cx,Dx) formula (0.40 pts)
  - Component 3: Column D formatted as date (yyyy-mm-dd), column E as integer (0.10 pts)
  - Component 4: Columns A, B, C, F are unchanged (0.10 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_networkdays_workdays_013'
SHEET_NAME = 'Review Schedule'
DATA_ROWS = range(2, 45)  # rows 2 to 44 inclusive (43 rows)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: D2:D44 contains =EDATE(Cx,6) formula (0.40 points)
    # This FAILS on initial (cells are empty) → PASSES on golden (cells have EDATE formulas)
    try:
        edate_correct = 0
        edate_total = 0
        edate_failures = []
        for row in DATA_ROWS:
            edate_total += 1
            d_val = ws.cell(row=row, column=4).value
            expected_formula = f'=EDATE(C{row},6)'
            if d_val is not None and isinstance(d_val, str):
                # Normalize for comparison: uppercase, no spaces
                d_norm = d_val.upper().replace(' ', '')
                exp_norm = expected_formula.upper().replace(' ', '')
                if d_norm == exp_norm:
                    edate_correct += 1
                else:
                    edate_failures.append(f"D{row}: expected {repr(expected_formula)}, got {repr(d_val)}")
            else:
                edate_failures.append(f"D{row}: expected formula, got {repr(d_val)}")

        if edate_correct == edate_total:
            print(f"PASS: Component 1 — D2:D44 all have =EDATE(Cx,6) formula ({edate_correct}/{edate_total} correct) (0.40 pts)")
            total_score += 0.40
        elif edate_correct > 0:
            # Partial credit for partial completion
            partial = round(0.40 * (edate_correct / edate_total), 4)
            print(f"PARTIAL: Component 1 — D2:D44 EDATE formulas: {edate_correct}/{edate_total} correct (+{partial} pts)")
            if edate_failures[:3]:
                print(f"  Sample failures: {edate_failures[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — D2:D44 EDATE formulas: {edate_correct}/{edate_total} correct (0.0 pts)")
            if edate_failures[:3]:
                print(f"  Sample failures: {edate_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — could not check D column EDATE formulas: {e}")

    # Component 2: E2:E44 contains =NETWORKDAYS(Cx,Dx) formula (0.40 points)
    # This FAILS on initial (cells are empty) → PASSES on golden (cells have NETWORKDAYS formulas)
    try:
        nd_correct = 0
        nd_total = 0
        nd_failures = []
        for row in DATA_ROWS:
            nd_total += 1
            e_val = ws.cell(row=row, column=5).value
            expected_formula = f'=NETWORKDAYS(C{row},D{row})'
            if e_val is not None and isinstance(e_val, str):
                # Normalize for comparison: uppercase, no spaces
                e_norm = e_val.upper().replace(' ', '')
                exp_norm = expected_formula.upper().replace(' ', '')
                if e_norm == exp_norm:
                    nd_correct += 1
                else:
                    nd_failures.append(f"E{row}: expected {repr(expected_formula)}, got {repr(e_val)}")
            else:
                nd_failures.append(f"E{row}: expected formula, got {repr(e_val)}")

        if nd_correct == nd_total:
            print(f"PASS: Component 2 — E2:E44 all have =NETWORKDAYS(Cx,Dx) formula ({nd_correct}/{nd_total} correct) (0.40 pts)")
            total_score += 0.40
        elif nd_correct > 0:
            partial = round(0.40 * (nd_correct / nd_total), 4)
            print(f"PARTIAL: Component 2 — E2:E44 NETWORKDAYS formulas: {nd_correct}/{nd_total} correct (+{partial} pts)")
            if nd_failures[:3]:
                print(f"  Sample failures: {nd_failures[:3]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — E2:E44 NETWORKDAYS formulas: {nd_correct}/{nd_total} correct (0.0 pts)")
            if nd_failures[:3]:
                print(f"  Sample failures: {nd_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — could not check E column NETWORKDAYS formulas: {e}")

    # Component 3: Column D formatted as date (yyyy-mm-dd), column E as integer '0' (0.10 points)
    # In the initial file, D2 and E2 are empty with 'General' format.
    # After task: D2 should have an explicit date format (e.g. 'yyyy-mm-dd'), E2 should be '0' (integer).
    # 'General' is NOT accepted here because empty cells in the initial file also have 'General'.
    try:
        d2_fmt = ws.cell(row=2, column=4).number_format
        e2_fmt = ws.cell(row=2, column=5).number_format

        # Accept common date formats for column D — must be explicit (not 'General')
        d_is_date = (d2_fmt.lower() != 'general' and
                     ('yy' in d2_fmt.lower() or 'mm' in d2_fmt.lower() or 'dd' in d2_fmt.lower()))
        # Accept only explicit integer format for column E — 'General' is NOT accepted
        e_is_int = e2_fmt in {'0', '#,##0', '0;0', '0_);(0)', '_(* #,##0_)'}

        if d_is_date and e_is_int:
            print(f"PASS: Component 3 — D2 format='{d2_fmt}' (date), E2 format='{e2_fmt}' (integer) (0.10 pts)")
            total_score += 0.10
        elif d_is_date:
            print(f"PARTIAL: Component 3 — D2 format='{d2_fmt}' (date OK), E2 format='{e2_fmt}' (expected integer like '0'). (+0.05 pts)")
            total_score += 0.05
        elif e_is_int:
            print(f"PARTIAL: Component 3 — D2 format='{d2_fmt}' (expected date format), E2 format='{e2_fmt}' (integer OK). (+0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — D2 format='{d2_fmt}' (expected date format like 'yyyy-mm-dd'), E2 format='{e2_fmt}' (expected integer '0') (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — could not check number formats: {e}")

    # Component 4: Columns A, B, C, F are unchanged (0.10 points)
    # Compare columns A/B/C/F in the tested file against the initial file.
    # IMPORTANT: The initial file path is hardcoded using TASK_ID so it is stable
    # regardless of whether we are testing golden or initial. When testing initial,
    # the comparison is against itself → always passes → but the component only awards
    # points when BOTH Component 1 AND Component 2 have non-zero scores (to prevent
    # the initial file from receiving these 0.10 points spuriously).
    # Better: We simply compare against the hardcoded initial file path.
    # Since the initial file always has identical A/B/C/F to itself, to make this
    # component fail on initial, we gate it: only award if at least Component 1 OR 2 passed.
    try:
        initial_path_hardcoded = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

        if os.path.exists(initial_path_hardcoded):
            wb_initial = openpyxl.load_workbook(initial_path_hardcoded)
            if SHEET_NAME in wb_initial.sheetnames:
                ws_initial = wb_initial[SHEET_NAME]
                unchanged = True
                violations = []
                preserved_cols = [1, 2, 3, 6]  # A, B, C, F
                for row in range(1, 45):
                    for col in preserved_cols:
                        v_i = ws_initial.cell(row=row, column=col).value
                        v_g = ws.cell(row=row, column=col).value
                        if v_i != v_g:
                            from openpyxl.utils import get_column_letter
                            violations.append(f"{get_column_letter(col)}{row}: {repr(v_i)} -> {repr(v_g)}")
                            unchanged = False

                # Gate: only award if some task-required change was detected (D or E has formulas)
                # This prevents initial file (which has identical A/B/C/F to itself) from scoring here
                d2_has_formula = ws.cell(row=2, column=4).value is not None
                e2_has_formula = ws.cell(row=2, column=5).value is not None
                task_changed = d2_has_formula or e2_has_formula

                if unchanged and task_changed:
                    print(f"PASS: Component 4 — Columns A/B/C/F unchanged from initial state (0.10 pts)")
                    total_score += 0.10
                elif not task_changed:
                    print(f"FAIL: Component 4 — No task changes detected in D/E columns; skipping preservation award (0.0 pts)")
                else:
                    print(f"FAIL: Component 4 — {len(violations)} cell(s) modified in preserved columns (0.0 pts)")
                    print(f"  Violations: {violations[:3]}")
            else:
                print(f"FAIL: Component 4 — initial file missing sheet '{SHEET_NAME}', cannot verify (0.0 pts)")
        else:
            print(f"SKIP: Component 4 — initial file not found at {initial_path_hardcoded} (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — could not check preserved columns: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
