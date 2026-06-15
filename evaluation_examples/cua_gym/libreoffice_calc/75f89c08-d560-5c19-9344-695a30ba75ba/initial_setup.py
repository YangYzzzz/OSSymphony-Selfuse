"""
Initial Setup: Split full names into First Name and Last Name columns
Task ID: calc_gg5_023
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


# 120 realistic full names in "LastName, FirstName" format
NAMES = [
    "Johnson, Michael", "Williams, Sarah", "Brown, James", "Garcia, Maria",
    "Martinez, David", "Anderson, Jennifer", "Taylor, Robert", "Thomas, Lisa",
    "Hernandez, Carlos", "Moore, Patricia", "Martin, Daniel", "Jackson, Nancy",
    "Thompson, Christopher", "White, Karen", "Lopez, Matthew", "Lee, Betty",
    "Gonzalez, Anthony", "Harris, Margaret", "Clark, Mark", "Robinson, Sandra",
    "Lewis, Steven", "Walker, Ashley", "Young, Kevin", "Allen, Dorothy",
    "King, Brian", "Wright, Kimberly", "Scott, Jason", "Torres, Emily",
    "Nguyen, Justin", "Hill, Donna", "Flores, Brandon", "Green, Carol",
    "Adams, Samuel", "Nelson, Michelle", "Baker, Gregory", "Hall, Amanda",
    "Rivera, Alexander", "Campbell, Melissa", "Mitchell, Patrick", "Carter, Deborah",
    "Roberts, Kenneth", "Gomez, Stephanie", "Phillips, Edward", "Evans, Rebecca",
    "Turner, Timothy", "Diaz, Laura", "Parker, Jose", "Cruz, Sharon",
    "Edwards, Nathan", "Collins, Cynthia", "Reyes, Frank", "Stewart, Kathleen",
    "Morris, Raymond", "Morales, Anna", "Murphy, Dennis", "Cook, Ruth",
    "Rogers, Jerry", "Gutierrez, Brenda", "Ortiz, Tyler", "Morgan, Pamela",
    "Cooper, Henry", "Peterson, Nicole", "Bailey, Douglas", "Reed, Samantha",
    "Kelly, Aaron", "Howard, Christine", "Ramos, Jack", "Cox, Catherine",
    "Ward, Albert", "Richardson, Frances", "Watson, Russell", "Brooks, Virginia",
    "Chavez, Roy", "Wood, Judith", "James, Eugene", "Bennett, Cheryl",
    "Gray, Philip", "Mendoza, Ann", "Ruiz, Bobby", "Hughes, Diana",
    "Price, Johnny", "Alvarez, Jean", "Castillo, Gerald", "Sanders, Kathryn",
    "Patel, Lawrence", "Myers, Joyce", "Long, Carl", "Foster, Grace",
    "Jimenez, Arthur", "Powell, Theresa", "Jenkins, Wayne", "Perry, Beverly",
    "Russell, Jesse", "Sullivan, Denise", "Bell, Roger", "Coleman, Tammy",
    "Butler, Ralph", "Henderson, Irene", "Barnes, Louis", "Simmons, Marilyn",
    "Patterson, Willie", "Bryant, Teresa", "Fisher, Billy", "Howard, Janice",
    "Ramirez, Christian", "Alexander, Judith", "Hayes, Austin", "Gibson, Rose",
    "Washington, Joe", "Burns, Julia", "Mason, Sean", "Griffin, Lillian",
    "Hunt, Dylan", "Warren, Victoria", "Dunn, Peter", "Olson, Megan",
    "Freeman, Corey", "Graham, Evelyn", "Dixon, Clarence", "Ford, Heather",
    "Garza, Ernest", "Ferguson, Danielle", "Wallace, Craig", "Weber, Jacqueline",
]

COMPANIES = [
    "Apex Dynamics", "BrightPath Solutions", "Cascade Analytics", "DataVault Corp",
    "EcoSphere Technologies", "Frontier Innovations", "GlobalReach Systems",
    "Horizon Digital", "Insight Partners", "JetStream Media", "KeyStone Industries",
    "Luminex Group", "MeridianTech", "NovaBridge Consulting", "Omni Solutions",
    "PrimeEdge Software", "Quantum Fields", "RedLeaf Enterprises", "SkyPeak Labs",
    "TrueNorth Advisors", "UrbanGrid Networks", "VantagePoint AI", "WaveSync Corp",
    "Xenith Ventures", "YieldMax Analytics", "ZenithCore Systems", "AlphaWave Digital",
    "BlueShift Robotics", "CrestView Partners", "DeltaForce Logistics",
]

CITIES = [
    "New York", "Los Angeles", "Chicago", "Houston", "Phoenix",
    "Philadelphia", "San Antonio", "San Diego", "Dallas", "San Jose",
    "Austin", "Jacksonville", "Fort Worth", "Columbus", "Charlotte",
    "Indianapolis", "San Francisco", "Seattle", "Denver", "Nashville",
    "Portland", "Oklahoma City", "Las Vegas", "Memphis", "Louisville",
    "Baltimore", "Milwaukee", "Albuquerque", "Tucson", "Sacramento",
]

STATUSES = ["Active", "Active", "Active", "Active", "Inactive", "Pending", "Active", "VIP", "Active", "Active"]

DOMAINS_EMAIL = [
    "gmail.com", "outlook.com", "yahoo.com", "company.com", "work.org",
    "mail.net", "proton.me", "icloud.com", "fastmail.com", "zoho.com",
]


def create_initial():
    wb = openpyxl.Workbook()

    # --- Contacts sheet ---
    ws = wb.active
    ws.title = 'Contacts'

    # Headers
    headers = ['Full Name', 'Email', 'Phone', 'Company', 'City', 'Status', 'First Name', 'Last Name']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 26
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    # 120 rows of data
    for i in range(120):
        row = i + 2
        full_name = NAMES[i]
        # Parse to build email
        parts = full_name.split(', ')
        last = parts[0]
        first = parts[1]
        email = f"{first.lower()}.{last.lower()}@{DOMAINS_EMAIL[i % len(DOMAINS_EMAIL)]}"
        phone = f"({200 + i % 800:03d}) {100 + (i * 37) % 900:03d}-{1000 + (i * 53) % 9000:04d}"
        company = COMPANIES[i % len(COMPANIES)]
        city = CITIES[i % len(CITIES)]
        status = STATUSES[i % len(STATUSES)]

        ws.cell(row=row, column=1, value=full_name)
        ws.cell(row=row, column=2, value=email)
        ws.cell(row=row, column=3, value=phone)
        ws.cell(row=row, column=4, value=company)
        ws.cell(row=row, column=5, value=city)
        ws.cell(row=row, column=6, value=status)
        # G and H are intentionally left EMPTY (task is to fill them with formulas)

    # Freeze header row
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H121'

    # --- CRM Notes sheet ---
    ws2 = wb.create_sheet('CRM Notes')
    ws2_headers = ['Date', 'Contact Name', 'Note', 'Follow-Up']
    for col, h in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    notes_data = [
        ['2025-03-01', 'Johnson, Michael', 'Discussed Q2 contract renewal', '2025-03-15'],
        ['2025-03-02', 'Williams, Sarah', 'Product demo scheduled', '2025-03-10'],
        ['2025-03-03', 'Garcia, Maria', 'Onboarding call completed', '2025-03-17'],
        ['2025-03-05', 'Anderson, Jennifer', 'Follow-up on support ticket #4521', '2025-03-12'],
        ['2025-03-07', 'Taylor, Robert', 'Annual review meeting', '2025-04-07'],
        ['2025-03-08', 'Moore, Patricia', 'Invoice inquiry resolved', '2025-03-22'],
        ['2025-03-10', 'Martin, Daniel', 'Upsell opportunity identified', '2025-03-24'],
        ['2025-03-11', 'White, Karen', 'Referral program enrollment', '2025-03-25'],
    ]
    for r, row_data in enumerate(notes_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # --- Campaign Stats sheet ---
    ws3 = wb.create_sheet('Campaign Stats')
    ws3_headers = ['Campaign', 'Emails Sent', 'Opens', 'Clicks', 'Conversions', 'Revenue']
    for col, h in enumerate(ws3_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    campaigns = [
        ['Spring Promo 2025', 4500, 1890, 432, 87, 34200],
        ['Product Launch Q1', 3200, 1504, 389, 112, 56700],
        ['Loyalty Rewards', 2800, 1232, 298, 64, 19800],
        ['Webinar Series', 1500, 945, 267, 53, 15900],
        ['Year-End Clearance', 5200, 2340, 578, 145, 72500],
    ]
    for r, row_data in enumerate(campaigns, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
