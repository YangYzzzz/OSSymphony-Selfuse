"""
Reward Script: Split full names into First Name and Last Name columns using text formulas
Task ID: calc_gg5_023
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): G column (First Name) has formulas in all 120 data rows
  - Component 2 (0.25): H column (Last Name) has formulas in all 120 data rows
  - Component 3 (0.25): G formulas use appropriate text functions (MID/RIGHT + FIND)
  - Component 4 (0.25): H formulas use appropriate text functions (LEFT + FIND)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Contacts' sheet exists
    if 'Contacts' not in wb.sheetnames:
        print("CRITICAL: 'Contacts' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contacts']

    # Precondition: Headers G1='First Name', H1='Last Name'
    g1 = ws.cell(1, 7).value
    h1 = ws.cell(1, 8).value
    if g1 != 'First Name' or h1 != 'Last Name':
        print(f"WARN: Headers - G1={g1}, H1={h1} (expected 'First Name', 'Last Name')")

    # Component 1: G column (First Name) has formulas in all 120 data rows (0.25 points)
    try:
        g_formula_count = 0
        g_nonempty_count = 0
        for r in range(2, 122):
            val = ws.cell(r, 7).value
            if val is not None:
                g_nonempty_count += 1
                if isinstance(val, str) and val.startswith('='):
                    g_formula_count += 1

        if g_formula_count >= 120:
            print(f"PASS: Component 1 - G column has formulas in all 120 rows ({g_formula_count}/120) (0.25 pts)")
            total_score += 0.25
        elif g_nonempty_count >= 120:
            # Has values but not formulas (maybe hardcoded) - partial credit
            print(f"PARTIAL: Component 1 - G column has values in {g_nonempty_count}/120 rows but only {g_formula_count} are formulas (0.10 pts)")
            total_score += 0.10
        elif g_formula_count > 0:
            # Some formulas present
            frac = g_formula_count / 120.0
            pts = round(0.25 * frac, 2)
            print(f"PARTIAL: Component 1 - G column has formulas in {g_formula_count}/120 rows ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 1 - G column has no formulas (non-empty: {g_nonempty_count}/120)")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: H column (Last Name) has formulas in all 120 data rows (0.25 points)
    try:
        h_formula_count = 0
        h_nonempty_count = 0
        for r in range(2, 122):
            val = ws.cell(r, 8).value
            if val is not None:
                h_nonempty_count += 1
                if isinstance(val, str) and val.startswith('='):
                    h_formula_count += 1

        if h_formula_count >= 120:
            print(f"PASS: Component 2 - H column has formulas in all 120 rows ({h_formula_count}/120) (0.25 pts)")
            total_score += 0.25
        elif h_nonempty_count >= 120:
            print(f"PARTIAL: Component 2 - H column has values in {h_nonempty_count}/120 rows but only {h_formula_count} are formulas (0.10 pts)")
            total_score += 0.10
        elif h_formula_count > 0:
            frac = h_formula_count / 120.0
            pts = round(0.25 * frac, 2)
            print(f"PARTIAL: Component 2 - H column has formulas in {h_formula_count}/120 rows ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 2 - H column has no formulas (non-empty: {h_nonempty_count}/120)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: G formulas use appropriate text functions for extracting first name (0.25 points)
    # Task requires MID/RIGHT + FIND (or equivalent text splitting approach)
    try:
        g_valid_formula_count = 0
        for r in range(2, 122):
            val = ws.cell(r, 7).value
            if isinstance(val, str) and val.startswith('='):
                upper_val = val.upper()
                # First name extraction should use FIND and one of MID/RIGHT
                uses_find = 'FIND' in upper_val or 'SEARCH' in upper_val
                uses_extract = 'MID' in upper_val or 'RIGHT' in upper_val
                if uses_find and uses_extract:
                    g_valid_formula_count += 1

        if g_valid_formula_count >= 120:
            print(f"PASS: Component 3 - All 120 G formulas use text functions (MID/RIGHT+FIND) (0.25 pts)")
            total_score += 0.25
        elif g_valid_formula_count >= 60:
            pts = round(0.25 * (g_valid_formula_count / 120.0), 2)
            print(f"PARTIAL: Component 3 - {g_valid_formula_count}/120 G formulas use proper text functions ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 3 - Only {g_valid_formula_count}/120 G formulas use text functions (MID/RIGHT+FIND)")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: H formulas use appropriate text functions for extracting last name (0.25 points)
    # Task requires LEFT + FIND (or equivalent text splitting approach)
    try:
        h_valid_formula_count = 0
        for r in range(2, 122):
            val = ws.cell(r, 8).value
            if isinstance(val, str) and val.startswith('='):
                upper_val = val.upper()
                # Last name extraction should use FIND and LEFT
                uses_find = 'FIND' in upper_val or 'SEARCH' in upper_val
                uses_extract = 'LEFT' in upper_val
                if uses_find and uses_extract:
                    h_valid_formula_count += 1

        if h_valid_formula_count >= 120:
            print(f"PASS: Component 4 - All 120 H formulas use text functions (LEFT+FIND) (0.25 pts)")
            total_score += 0.25
        elif h_valid_formula_count >= 60:
            pts = round(0.25 * (h_valid_formula_count / 120.0), 2)
            print(f"PARTIAL: Component 4 - {h_valid_formula_count}/120 H formulas use proper text functions ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 4 - Only {h_valid_formula_count}/120 H formulas use text functions (LEFT+FIND)")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
