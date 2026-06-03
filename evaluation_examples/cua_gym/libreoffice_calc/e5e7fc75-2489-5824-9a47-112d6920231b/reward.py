"""
FINAL REWARD SCRIPT - SUCCESS
Task: Calculate the error rate in a new column titled "Error Rate (%)", ensure number format, and mark the largest value with dark red (#cc0000) font.
Generated: 2025-11-24 07:27:06
Status: success
Model: o3
Total Steps: 6
"""

import openpyxl
import os
import math


def verify_error_rate_task(file_path: str) -> float:
    """Verify that the spreadsheet meets all task requirements:
    1. New column titled "Error Rate (%)" exists
    2. Each row correctly calculates error rate (Errors / Total) – either via a correct
       formula or a numeric value matching the expected result
    3. Cells in the new column use a percentage number format
    4. The largest error-rate value is displayed in dark-red font (#cc0000)

    Returns a progressive score between 0.0 and 1.0
    """

    print(f"Starting verification for: {file_path}")

    # ---------- Early checks ----------
    if not os.path.isfile(file_path):
        print("✗ File not found")
        return 0.0

    # Weights for progressive scoring (must sum to 1.0)
    W_COLUMN = 0.20   # column present
    W_VALUES = 0.35   # calculation / formula correctness
    W_FORMAT = 0.15   # percentage number format
    W_COLOR  = 0.30   # largest value dark-red font

    score = 0.0  # progressive score accumulator

    # Load workbook twice: once with data_only=True for computed values,
    # once normally to inspect formulas & styles
    try:
        wb_values = openpyxl.load_workbook(file_path, data_only=True)
        wb_styles = openpyxl.load_workbook(file_path, data_only=False)
        print("✓ Workbook loaded successfully (values & styles)")
    except Exception as e:
        print(f"✗ Unable to load workbook: {e}")
        return 0.0

    sh_values = wb_values.active
    sh_styles = wb_styles[sh_values.title]

    # ---------- 1. Column existence ----------
    headers = [cell.value for cell in sh_values[1]]
    print("Headers detected:", headers)
    try:
        err_col_idx = headers.index("Error Rate (%)") + 1  # 1-based index
        print(f"✓ 'Error Rate (%)' column found at index {err_col_idx}")
        score += W_COLUMN
    except ValueError:
        print("✗ Required column 'Error Rate (%)' is missing")
        return score  # cannot continue meaningful verification

    # ---------- 2 & 3. Row-by-row checks ----------
    all_calc_correct = True
    all_percent_fmt  = True
    expected_rates   = []  # store expected numeric rates for later colour check

    row = 2
    while True:
        id_val    = sh_values.cell(row=row, column=1).value
        total_val = sh_values.cell(row=row, column=2).value
        err_val   = sh_values.cell(row=row, column=3).value

        # Terminate when we hit an empty data row (all key fields None)
        if id_val is None and total_val is None and err_val is None:
            break

        # Skip rows lacking numeric data
        try:
            total_num = float(total_val)
            err_num   = float(err_val)
        except (TypeError, ValueError):
            print(f"✗ Row {row}: Non-numeric 'Total' or 'Errors'")
            all_calc_correct = False
            row += 1
            continue

        expected_rate = err_num / total_num if total_num != 0 else None
        expected_rates.append(expected_rate)

        # ----- 2a. Formula / value correctness -----
        cell_style = sh_styles.cell(row=row, column=err_col_idx)
        formula_or_value = cell_style.value  # may be formula string or literal
        data_only_value  = sh_values.cell(row=row, column=err_col_idx).value
        calc_ok = False

        # Accept either a formula that references C{row}/B{row}, or a numeric value
        if isinstance(formula_or_value, str) and formula_or_value.startswith('='):
            cleaned = formula_or_value.upper().replace('$', '')
            if f"C{row}" in cleaned and f"B{row}" in cleaned and '/' in cleaned:
                calc_ok = True
        else:
            try:
                if math.isclose(float(data_only_value), expected_rate, rel_tol=1e-4, abs_tol=1e-4):
                    calc_ok = True
            except (TypeError, ValueError):
                pass

        if calc_ok:
            print(f"✓ Row {row}: calculation/formula correct")
        else:
            print(f"✗ Row {row}: incorrect calculation or formula (cell='{formula_or_value}')")
            all_calc_correct = False

        # ----- 3. Percentage number format -----
        number_format = cell_style.number_format
        if '%' not in str(number_format):
            all_percent_fmt = False

        row += 1

    if not expected_rates:
        print("✗ No data rows found for verification")
        return score

    if all_calc_correct:
        print("✓ All error-rate calculations/formulas are correct")
        score += W_VALUES
    else:
        print("✗ Some error-rate calculations/formulas are incorrect")

    if all_percent_fmt:
        print("✓ Percentage number format applied to all error-rate cells")
        score += W_FORMAT
    else:
        print("✗ Percentage format missing or incorrect in some cells")

    # ---------- 4. Dark-red colour for largest value ----------
    max_rate = max(expected_rates)
    DARK_RED_RGB = 'FFCC0000'  # openpyxl stores RGB with alpha prefix
    dark_red_found = False

    for idx, expected in enumerate(expected_rates, start=2):
        if math.isclose(expected, max_rate, rel_tol=1e-12, abs_tol=1e-12):
            font_color = sh_styles.cell(row=idx, column=err_col_idx).font.color
            rgb_val = None
            if font_color is not None:
                rgb_val = getattr(font_color, 'rgb', None) if font_color.type == 'rgb' else font_color.rgb
            print(f"Row {idx}: font colour detected {rgb_val}")
            if rgb_val and rgb_val.upper().endswith('CC0000'):
                dark_red_found = True

    if dark_red_found:
        print("✓ Largest error-rate value is coloured dark red (#cc0000)")
        score += W_COLOR
    else:
        print("✗ Largest error-rate value is not coloured dark red (#cc0000)")

    # ---------- Final score ----------
    final_score = min(score, 1.0)
    print(f"Final score: {final_score}")
    print(f"REWARD: {final_score}")
    return final_score


# --------------------
# Execute verification
# --------------------
if __name__ == "__main__":
    FILE_PATH = "/home/user/calculate_the_error_rate_in_a_new_column_titled_error_rate_ensure_number_format_and_mark_the_largest.xlsx"
    verify_error_rate_task(FILE_PATH)
