"""
Reward Script: Define named range 'Inventory_Qty' and SUMIF formula in G2
Task ID: calc_nrv_026
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Named range 'Inventory_Qty' exists with correct reference
  Component 2 (0.3): Cell G2 contains a SUMIF formula
  Component 3 (0.3): G2 formula references 'In Stock' criteria and Inventory_Qty named range
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_026'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Warehouse' sheet must exist
    if 'Warehouse' not in wb.sheetnames:
        print("CRITICAL: 'Warehouse' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Warehouse']

    # Component 1: Named range 'Inventory_Qty' exists and refers to Warehouse!$C$2:$C$200 (0.4 points)
    try:
        matching_names = [dn for dn in wb.defined_names.values()
                         if dn.name.lower() == 'inventory_qty']
        if len(matching_names) > 0:
            ref = matching_names[0].attr_text
            # Normalize: remove quotes around sheet name if present
            normalized_ref = ref.replace("'", "").upper().replace(" ", "")
            expected_ref = "WAREHOUSE!$C$2:$C$200"
            if normalized_ref == expected_ref:
                print(f"PASS: Component 1 — Named range 'Inventory_Qty' = {ref} (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 1 — Named range 'Inventory_Qty' exists but ref is '{ref}', expected 'Warehouse!$C$2:$C$200'")
        else:
            print("FAIL: Component 1 — Named range 'Inventory_Qty' not found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cell G2 contains a SUMIF formula (0.3 points)
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str) and 'SUMIF' in g2_value.upper():
            print(f"PASS: Component 2 — G2 contains SUMIF formula: {g2_value} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — G2 expected SUMIF formula, found: {g2_value}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2 formula references 'In Stock' criteria and uses Inventory_Qty named range (0.3 points)
    try:
        g2_value = ws['G2'].value
        if g2_value is not None and isinstance(g2_value, str):
            formula_upper = g2_value.upper().replace(" ", "")
            has_in_stock = '"INSTOCK"' in formula_upper or "'INSTOCK'" in formula_upper or '"IN STOCK"' in g2_value or "'In Stock'" in g2_value
            # Also check the original with spaces preserved
            if not has_in_stock:
                has_in_stock = 'IN STOCK' in g2_value.upper() and ('"' in g2_value or "'" in g2_value)
            has_named_range = 'INVENTORY_QTY' in formula_upper
            if has_in_stock and has_named_range:
                print(f"PASS: Component 3 — G2 formula uses 'In Stock' criteria and Inventory_Qty range (0.3 pts)")
                total_score += 0.3
            elif has_in_stock:
                print(f"FAIL: Component 3 — G2 has 'In Stock' criteria but does not reference Inventory_Qty")
            elif has_named_range:
                print(f"FAIL: Component 3 — G2 references Inventory_Qty but missing 'In Stock' criteria")
            else:
                print(f"FAIL: Component 3 — G2 formula missing both 'In Stock' criteria and Inventory_Qty: {g2_value}")
        else:
            print(f"FAIL: Component 3 — G2 is not a formula: {g2_value}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
