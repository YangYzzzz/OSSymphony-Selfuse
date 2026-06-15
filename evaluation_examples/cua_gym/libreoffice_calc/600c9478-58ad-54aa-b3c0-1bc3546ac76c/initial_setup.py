"""
Initial Setup: Define named range and SUMIF formula for warehouse inventory
Task ID: calc_nrv_026
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Warehouse'

    # Headers
    headers = ['SKU', 'Item', 'Quantity', 'Location', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # G1 header
    ws['G1'] = 'In Stock Total'
    # G2 intentionally left empty (task requires agent to fill it)

    # Generate 199 rows of realistic inventory data
    item_prefixes = [
        'Industrial', 'Heavy-Duty', 'Premium', 'Standard', 'Economy',
        'Professional', 'Commercial', 'Compact', 'Ultra', 'Max'
    ]
    item_bases = [
        'Bearing', 'Gasket', 'Valve', 'Coupling', 'Flange',
        'Bracket', 'Bushing', 'Fastener', 'Seal', 'Hinge',
        'Clamp', 'Spacer', 'Washer', 'Spring', 'Pulley',
        'Shaft', 'Gear', 'Sprocket', 'Belt', 'Filter'
    ]
    item_suffixes = [
        'Assembly', 'Kit', 'Set', 'Unit', 'Pack',
        'Module', 'Component', 'Block', 'Ring', 'Plate'
    ]
    locations = [
        'Aisle-A Rack-1', 'Aisle-A Rack-2', 'Aisle-A Rack-3',
        'Aisle-B Rack-1', 'Aisle-B Rack-2', 'Aisle-B Rack-3',
        'Aisle-C Rack-1', 'Aisle-C Rack-2', 'Aisle-C Rack-3',
        'Aisle-D Rack-1', 'Aisle-D Rack-2', 'Aisle-D Rack-3',
        'Aisle-E Rack-1', 'Aisle-E Rack-2', 'Aisle-E Rack-3',
        'Bay-1 Shelf-A', 'Bay-1 Shelf-B', 'Bay-2 Shelf-A',
        'Bay-2 Shelf-B', 'Bay-3 Shelf-A', 'Bay-3 Shelf-B',
        'Mezzanine-1', 'Mezzanine-2', 'Dock-A', 'Dock-B'
    ]
    statuses = ['In Stock', 'Back Order', 'Discontinued']
    status_weights = [0.60, 0.25, 0.15]

    for i in range(1, 200):
        row = i + 1
        # SKU: WH-XXXXX
        sku = f'WH-{10000 + i:05d}'
        # Item name
        prefix = random.choice(item_prefixes)
        base = random.choice(item_bases)
        suffix = random.choice(item_suffixes)
        item_name = f'{prefix} {base} {suffix}'
        # Quantity: 1-500
        qty = random.randint(1, 500)
        # Location
        loc = random.choice(locations)
        # Status with weighted distribution
        status = random.choices(statuses, weights=status_weights, k=1)[0]

        ws.cell(row=row, column=1, value=sku)
        ws.cell(row=row, column=2, value=item_name)
        ws.cell(row=row, column=3, value=qty)
        ws.cell(row=row, column=4, value=loc)
        ws.cell(row=row, column=5, value=status)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['G'].width = 16

    # NO named ranges (task requires agent to create them)
    # NO formula in G2 (task requires agent to add it)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
