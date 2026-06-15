"""
Initial Setup: Apply duplicate-values conditional formatting to email column
Task ID: calc_ggf_012
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    # Headers
    headers = ['Email', 'First Name', 'Last Name', 'Company']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 100 rows of realistic contact data with some deliberate duplicate emails
    contacts = [
        ('sarah.chen@techvault.io', 'Sarah', 'Chen', 'TechVault Inc.'),
        ('marcus.johnson@blueridge.com', 'Marcus', 'Johnson', 'BlueRidge Corp'),
        ('elena.rodriguez@synthwave.co', 'Elena', 'Rodriguez', 'Synthwave Labs'),
        ('james.murphy@greenleaf.org', 'James', 'Murphy', 'GreenLeaf Foundation'),
        ('aisha.patel@orbitaldata.com', 'Aisha', 'Patel', 'Orbital Data Systems'),
        ('david.kim@nextera.io', 'David', 'Kim', 'NextEra Technologies'),
        ('rachel.foster@cloudpeak.net', 'Rachel', 'Foster', 'CloudPeak Networks'),
        ('thomas.wright@steelbridge.com', 'Thomas', 'Wright', 'SteelBridge Manufacturing'),
        ('maria.gonzalez@vivacreative.co', 'Maria', 'Gonzalez', 'Viva Creative Agency'),
        ('ryan.taylor@quantumleap.io', 'Ryan', 'Taylor', 'QuantumLeap Software'),
        # Duplicate of row 2 - sarah.chen
        ('sarah.chen@techvault.io', 'Sarah', 'Chen', 'TechVault Inc.'),
        ('olivia.brown@harborview.com', 'Olivia', 'Brown', 'HarborView Consulting'),
        ('nathan.lee@firestorm.dev', 'Nathan', 'Lee', 'Firestorm Development'),
        ('julia.martinez@solarpulse.com', 'Julia', 'Martinez', 'SolarPulse Energy'),
        ('kevin.anderson@ironclad.io', 'Kevin', 'Anderson', 'Ironclad Security'),
        ('samantha.davis@peakpoint.net', 'Samantha', 'Davis', 'PeakPoint Analytics'),
        ('christopher.white@titanforge.com', 'Christopher', 'White', 'TitanForge Industries'),
        ('michelle.thompson@brightpath.org', 'Michelle', 'Thompson', 'BrightPath Education'),
        ('daniel.garcia@corevolt.io', 'Daniel', 'Garcia', 'CoreVolt Systems'),
        ('jessica.moore@clearwater.co', 'Jessica', 'Moore', 'Clearwater Solutions'),
        # Duplicate of row 6 - david.kim
        ('david.kim@nextera.io', 'David', 'Kim', 'NextEra Technologies'),
        ('amanda.jackson@ridgeline.com', 'Amanda', 'Jackson', 'Ridgeline Partners'),
        ('matthew.harris@deepblue.net', 'Matthew', 'Harris', 'DeepBlue Research'),
        ('lauren.clark@suncoast.io', 'Lauren', 'Clark', 'Suncoast Digital'),
        ('brandon.lewis@apexgrowth.com', 'Brandon', 'Lewis', 'Apex Growth Capital'),
        ('stephanie.robinson@northwind.co', 'Stephanie', 'Robinson', 'Northwind Logistics'),
        ('andrew.walker@brightlabs.io', 'Andrew', 'Walker', 'Bright Labs'),
        ('megan.hall@vortexmedia.com', 'Megan', 'Hall', 'Vortex Media Group'),
        ('jonathan.allen@pinnaclehr.net', 'Jonathan', 'Allen', 'Pinnacle HR'),
        ('ashley.young@coastaltech.co', 'Ashley', 'Young', 'Coastal Technologies'),
        # Duplicate of row 10 - ryan.taylor
        ('ryan.taylor@quantumleap.io', 'Ryan', 'Taylor', 'QuantumLeap Software'),
        ('nicole.king@stratosphere.io', 'Nicole', 'King', 'Stratosphere Ventures'),
        ('justin.hernandez@bluecrest.com', 'Justin', 'Hernandez', 'BlueCrest Financial'),
        ('emily.scott@evergreensys.net', 'Emily', 'Scott', 'Evergreen Systems'),
        ('adam.green@novatech.co', 'Adam', 'Green', 'NovaTech Solutions'),
        ('heather.adams@goldengate.io', 'Heather', 'Adams', 'GoldenGate Innovations'),
        ('patrick.baker@redwood.com', 'Patrick', 'Baker', 'Redwood Analytics'),
        ('victoria.nelson@crystalclear.net', 'Victoria', 'Nelson', 'Crystal Clear Optics'),
        ('benjamin.carter@summitpro.io', 'Benjamin', 'Carter', 'SummitPro Consulting'),
        ('katherine.mitchell@bluepoint.co', 'Katherine', 'Mitchell', 'BluePoint Strategy'),
        ('derek.perez@thunderbolt.com', 'Derek', 'Perez', 'Thunderbolt Electric'),
        ('jennifer.roberts@horizonline.io', 'Jennifer', 'Roberts', 'HorizonLine Designs'),
        # Duplicate of row 5 - aisha.patel
        ('aisha.patel@orbitaldata.com', 'Aisha', 'Patel', 'Orbital Data Systems'),
        ('tyler.turner@maplesoft.net', 'Tyler', 'Turner', 'MapleSoft Technologies'),
        ('rebecca.phillips@silveroak.co', 'Rebecca', 'Phillips', 'Silver Oak Ventures'),
        ('william.campbell@gridlock.io', 'William', 'Campbell', 'GridLock Infrastructure'),
        ('danielle.parker@brightedge.com', 'Danielle', 'Parker', 'BrightEdge Marketing'),
        ('gregory.evans@forgepoint.net', 'Gregory', 'Evans', 'ForgePoint Capital'),
        ('allison.edwards@clearsky.io', 'Allison', 'Edwards', 'ClearSky Aviation'),
        ('marcus.collins@vanguardai.co', 'Marcus', 'Collins', 'Vanguard AI'),
        ('lisa.stewart@oceancrest.com', 'Lisa', 'Stewart', 'OceanCrest Resort'),
        ('eric.sanchez@ironpeak.io', 'Eric', 'Sanchez', 'IronPeak Mining'),
        ('tiffany.morris@starlightmed.net', 'Tiffany', 'Morris', 'Starlight Medical'),
        ('joseph.rogers@nexuswave.co', 'Joseph', 'Rogers', 'NexusWave Telecom'),
        ('christina.reed@alpineview.com', 'Christina', 'Reed', 'AlpineView Properties'),
        ('alexander.cook@pulsedrive.io', 'Alexander', 'Cook', 'PulseDrive Motors'),
        ('brittany.morgan@sapphiretech.net', 'Brittany', 'Morgan', 'Sapphire Technologies'),
        ('jason.bell@vertexdata.co', 'Jason', 'Bell', 'Vertex Data Corp'),
        ('laura.murphy@prismworks.io', 'Laura', 'Murphy', 'PrismWorks Studio'),
        ('sean.bailey@crestview.com', 'Sean', 'Bailey', 'Crestview Financial'),
        # Duplicate of row 3 - elena.rodriguez
        ('elena.rodriguez@synthwave.co', 'Elena', 'Rodriguez', 'Synthwave Labs'),
        ('diana.rivera@solarflare.net', 'Diana', 'Rivera', 'SolarFlare Networks'),
        ('charles.cooper@eaglecrest.io', 'Charles', 'Cooper', 'EagleCrest Defense'),
        ('anna.richardson@mosaiclab.co', 'Anna', 'Richardson', 'Mosaic Lab'),
        ('robert.cox@titaniumcore.com', 'Robert', 'Cox', 'TitaniumCore Systems'),
        ('vanessa.howard@auroratech.io', 'Vanessa', 'Howard', 'Aurora Technologies'),
        ('peter.ward@blackstone.net', 'Peter', 'Ward', 'Blackstone Engineering'),
        ('kelly.torres@riveredge.co', 'Kelly', 'Torres', 'RiverEdge Capital'),
        ('george.peterson@skylinepro.com', 'George', 'Peterson', 'Skyline Professional'),
        ('sandra.gray@diamondwave.io', 'Sandra', 'Gray', 'DiamondWave Media'),
        ('frank.ramirez@ironhorse.net', 'Frank', 'Ramirez', 'IronHorse Transport'),
        ('tracy.james@emberhub.co', 'Tracy', 'James', 'EmberHub Creative'),
        ('steven.watson@polaristech.com', 'Steven', 'Watson', 'Polaris Technologies'),
        ('nancy.brooks@coppervine.io', 'Nancy', 'Brooks', 'CopperVine Wines'),
        ('mark.kelly@summitline.net', 'Mark', 'Kelly', 'SummitLine Engineering'),
        ('carol.price@northstar.co', 'Carol', 'Price', 'NorthStar Navigation'),
        ('brian.bennett@deeprock.com', 'Brian', 'Bennett', 'DeepRock Exploration'),
        ('susan.wood@fleetwave.io', 'Susan', 'Wood', 'FleetWave Logistics'),
        ('keith.barnes@spirepoint.net', 'Keith', 'Barnes', 'SpirePoint Architecture'),
        ('donna.ross@clearpath.co', 'Donna', 'Ross', 'ClearPath Robotics'),
        ('timothy.henderson@bluevault.com', 'Timothy', 'Henderson', 'BlueVault Security'),
        ('janet.coleman@greenfield.io', 'Janet', 'Coleman', 'GreenField Agriculture'),
        ('ralph.jenkins@titanworks.net', 'Ralph', 'Jenkins', 'TitanWorks Manufacturing'),
        ('pamela.perry@suncrest.co', 'Pamela', 'Perry', 'SunCrest Solar'),
        ('harold.powell@meridianlab.com', 'Harold', 'Powell', 'Meridian Lab'),
        ('martha.long@edgepoint.io', 'Martha', 'Long', 'EdgePoint Analytics'),
        ('henry.patterson@falconridge.net', 'Henry', 'Patterson', 'FalconRidge Aviation'),
        ('gloria.hughes@coastalbreeze.co', 'Gloria', 'Hughes', 'Coastal Breeze Travel'),
        ('eugene.flores@rocksolid.com', 'Eugene', 'Flores', 'RockSolid Construction'),
        ('ruth.washington@primecore.io', 'Ruth', 'Washington', 'PrimeCore Data'),
        ('albert.butler@windmill.net', 'Albert', 'Butler', 'Windmill Energy'),
        ('sharon.simmons@blazetrail.co', 'Sharon', 'Simmons', 'BlazeTrail Adventures'),
        ('howard.foster@deepcurrent.com', 'Howard', 'Foster', 'DeepCurrent Research'),
        ('virginia.gonzales@skyward.io', 'Virginia', 'Gonzales', 'Skyward Aerospace'),
        ('carl.bryant@ironclad.net', 'Carl', 'Bryant', 'Ironclad Defense'),
        ('dorothy.alexander@quartzlabs.co', 'Dorothy', 'Alexander', 'Quartz Labs'),
        ('philip.russell@stormbreak.com', 'Philip', 'Russell', 'StormBreak Weather'),
        ('norma.griffin@goldenvale.io', 'Norma', 'Griffin', 'GoldenVale Estates'),
        ('wayne.diaz@crestfield.com', 'Wayne', 'Diaz', 'Crestfield Partners'),
        ('linda.sanders@horizonbay.io', 'Linda', 'Sanders', 'HorizonBay Shipping'),
    ]

    assert len(contacts) == 100, f"Expected 100 contacts, got {len(contacts)}"

    for r, (email, first, last, company) in enumerate(contacts, 2):
        ws.cell(row=r, column=1, value=email)
        ws.cell(row=r, column=2, value=first)
        ws.cell(row=r, column=3, value=last)
        ws.cell(row=r, column=4, value=company)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

    # NO conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
