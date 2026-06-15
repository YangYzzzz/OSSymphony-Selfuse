"""
Reward Script: Supply Chain Scorecard with KPI scores, weighted scores, and performance index
Task ID: calc_ops_094
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): F2:F9 Score formulas with correct direction (lower-is-better vs higher-is-better)
  Component 2 (0.3): G2:G9 Weighted Score formulas (Weight * Score)
  Component 3 (0.3): B11 Supply Chain Index = SUM of weighted scores
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_094'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active
    if ws is None:
        # Try by name
        if 'Scorecard' in wb.sheetnames:
            ws = wb['Scorecard']
        else:
            ws = wb.worksheets[0]

    # --- Component 1: F2:F9 Score formulas (0.4 points) ---
    # For rows where lower is better (3=Logistics Cost, 4=Defect Rate, 7=Lead Time): Score = MIN(D/E, 1)
    # For rows where higher is better (2,5,6,8,9): Score = MIN(E/D, 1)
    try:
        # Lower-is-better KPIs: Logistics Cost % (3), Defect Rate (4), Lead Time (7), New Product Time (9)
        # For these, Score = Target/Actual = D/E (lower actual is better)
        # Higher-is-better KPIs: COGS % Revenue (2), Supplier Quality (5), OTIF % (6), Order Change % (8)
        # For these, Score = Actual/Target = E/D (higher actual is better)
        lower_is_better_rows = {3, 4, 7, 9}
        higher_is_better_rows = {2, 5, 6, 8}

        f_correct = 0
        f_total = 8  # F2 through F9

        for row in range(2, 10):
            f_val = ws.cell(row=row, column=6).value  # Column F
            f_norm = normalize_formula(f_val)

            if row in lower_is_better_rows:
                # Expected: =MIN(D{row}/E{row},1) — Target/Actual
                expected_pattern = f'=MIN(D{row}/E{row},1)'
                expected_norm = normalize_formula(expected_pattern)
                if f_norm == expected_norm:
                    f_correct += 1
                    print(f"  PASS: F{row} correct lower-is-better formula: {f_val}")
                else:
                    # Also accept without MIN wrapper or slight variants
                    # Check if it's at least a formula involving D/E division
                    if f_norm and f'D{row}/E{row}' in f_norm:
                        f_correct += 0.5
                        print(f"  PARTIAL: F{row} has D/E division but not exact: {f_val}")
                    else:
                        print(f"  FAIL: F{row} expected {expected_pattern}, found: {f_val!r}")
            else:
                # Expected: =MIN(E{row}/D{row},1) — Actual/Target
                expected_pattern = f'=MIN(E{row}/D{row},1)'
                expected_norm = normalize_formula(expected_pattern)
                if f_norm == expected_norm:
                    f_correct += 1
                    print(f"  PASS: F{row} correct higher-is-better formula: {f_val}")
                else:
                    if f_norm and f'E{row}/D{row}' in f_norm:
                        f_correct += 0.5
                        print(f"  PARTIAL: F{row} has E/D division but not exact: {f_val}")
                    else:
                        print(f"  FAIL: F{row} expected {expected_pattern}, found: {f_val!r}")

        comp1_score = (f_correct / f_total) * 0.4
        if comp1_score > 0:
            print(f"PASS: Component 1 — F column Score formulas: {f_correct}/{f_total} correct ({comp1_score:.3f} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 — No correct Score formulas in F2:F9")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: G2:G9 Weighted Score formulas (0.3 points) ---
    # Expected: =C{row}*F{row} for each row 2-9
    try:
        g_correct = 0
        g_total = 8  # G2 through G9

        for row in range(2, 10):
            g_val = ws.cell(row=row, column=7).value  # Column G
            g_norm = normalize_formula(g_val)

            expected_pattern = f'=C{row}*F{row}'
            expected_norm = normalize_formula(expected_pattern)

            # Also accept =F{row}*C{row}
            alt_pattern = f'=F{row}*C{row}'
            alt_norm = normalize_formula(alt_pattern)

            if g_norm == expected_norm or g_norm == alt_norm:
                g_correct += 1
                print(f"  PASS: G{row} correct weighted score formula: {g_val}")
            else:
                # Check for any formula that references both C and F for this row
                if g_norm and f'C{row}' in g_norm and f'F{row}' in g_norm:
                    g_correct += 0.5
                    print(f"  PARTIAL: G{row} references C and F but not exact: {g_val}")
                else:
                    print(f"  FAIL: G{row} expected {expected_pattern}, found: {g_val!r}")

        comp2_score = (g_correct / g_total) * 0.3
        if comp2_score > 0:
            print(f"PASS: Component 2 — G column Weighted Score formulas: {g_correct}/{g_total} correct ({comp2_score:.3f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — No correct Weighted Score formulas in G2:G9")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: B11 Supply Chain Index formula (0.3 points) ---
    # Expected: =SUM(G2:G9)
    try:
        b11_val = ws.cell(row=11, column=2).value
        b11_norm = normalize_formula(b11_val)

        expected_sum = normalize_formula('=SUM(G2:G9)')

        if b11_norm == expected_sum:
            print(f"PASS: Component 3 — B11 Supply Chain Index formula correct: {b11_val} (0.3 pts)")
            total_score += 0.3
        else:
            # Accept any SUM formula referencing G column range
            if b11_norm and 'SUM' in b11_norm and 'G' in b11_norm:
                print(f"PARTIAL: Component 3 — B11 has SUM formula with G reference: {b11_val} (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 3 — B11 expected =SUM(G2:G9), found: {b11_val!r}")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook for LibreOffice — save any unsaved changes before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
