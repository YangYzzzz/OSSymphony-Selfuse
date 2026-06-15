"""
Initial Setup: Supply Chain Scorecard with KPI data (no formulas)
Task ID: calc_ops_094
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Scorecard Sheet ---
    ws = wb.active
    ws.title = 'Scorecard'

    # Headers
    headers = ['Perspective', 'KPI', 'Weight', 'Target', 'Actual',
               'Score (Actual/Target)', 'Weighted Score']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows (raw data only, NO formulas in F or G columns)
    data = [
        ['Cost',        'COGS % Revenue',     0.15, 0.60, 0.58],
        ['Cost',        'Logistics Cost %',   0.10, 0.05, 0.06],
        ['Quality',     'Defect Rate',        0.15, 0.01, 0.008],
        ['Quality',     'Supplier Quality',   0.10, 0.98, 0.97],
        ['Delivery',    'OTIF %',             0.20, 0.95, 0.93],
        ['Delivery',    'Lead Time (days)',   0.10, 5,    6],
        ['Flexibility', 'Order Change %',     0.10, 0.90, 0.85],
        ['Flexibility', 'New Product Time',   0.10, 30,   28],
    ]

    data_font = Font(name='Calibri', size=11)
    data_align = Alignment(horizontal='center', vertical='center')

    # Alternating row fills for readability
    light_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    for r, row_data in enumerate(data, 2):
        row_fill = light_fill if r % 2 == 0 else white_fill
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border
            cell.fill = row_fill

        # Format Weight column as percentage
        ws.cell(row=r, column=3).number_format = '0%'

        # Format Target and Actual based on row type
        if row_data[1] in ('COGS % Revenue', 'Logistics Cost %', 'Defect Rate',
                           'Supplier Quality', 'OTIF %', 'Order Change %'):
            ws.cell(row=r, column=4).number_format = '0.00%'
            ws.cell(row=r, column=5).number_format = '0.00%'
        else:
            ws.cell(row=r, column=4).number_format = '0'
            ws.cell(row=r, column=5).number_format = '0'

        # F and G columns are intentionally empty (task is to add formulas)
        # Leave them blank but styled
        for col in [6, 7]:
            cell = ws.cell(row=r, column=col)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = data_align

    # Row 10 is empty (separator)

    # Row 11: Supply Chain Index label
    ws.cell(row=11, column=1, value='Supply Chain Index')
    ws['A11'].font = Font(name='Calibri', size=12, bold=True, color='2F5496')
    ws['A11'].alignment = Alignment(horizontal='left', vertical='center')
    # B11 is intentionally empty (task is to put the sum here)

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
