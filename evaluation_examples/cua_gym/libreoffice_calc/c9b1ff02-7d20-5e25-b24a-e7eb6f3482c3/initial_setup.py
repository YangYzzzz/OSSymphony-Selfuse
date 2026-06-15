"""
Initial Setup: Create Task_Completion spreadsheet with project completion data
Task ID: calc_gcv_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Task_Completion"

    # --- Headers ---
    headers = ["Project", "Phase", "Task", "Owner", "Target", "Completion %"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    # --- Data: 34 rows of realistic project tracking data ---
    data = [
        # Project, Phase, Task, Owner, Target, Completion %
        ["ERP Migration", "Planning", "Requirements gathering", "Sarah Chen", "2025-03-15", 0.92],
        ["ERP Migration", "Planning", "Vendor evaluation report", "Sarah Chen", "2025-03-20", 0.85],
        ["ERP Migration", "Planning", "Budget approval documentation", "Marcus Johnson", "2025-03-25", 1.00],
        ["ERP Migration", "Design", "System architecture diagram", "Raj Patel", "2025-04-10", 0.68],
        ["ERP Migration", "Design", "Data mapping specification", "Raj Patel", "2025-04-15", 0.55],
        ["ERP Migration", "Development", "Core module integration", "Lena Kowalski", "2025-05-01", 0.42],
        ["ERP Migration", "Development", "API endpoint development", "Lena Kowalski", "2025-05-15", 0.30],
        ["ERP Migration", "Testing", "Unit test execution", "Tom Rivera", "2025-06-01", 0.15],
        ["Website Redesign", "Planning", "Stakeholder interviews", "Aisha Mohammed", "2025-03-10", 0.95],
        ["Website Redesign", "Planning", "Competitive analysis report", "Aisha Mohammed", "2025-03-18", 0.78],
        ["Website Redesign", "Design", "Wireframe creation", "Wei Zhang", "2025-04-05", 0.62],
        ["Website Redesign", "Design", "Style guide finalization", "Wei Zhang", "2025-04-12", 0.50],
        ["Website Redesign", "Development", "Frontend component library", "Carlos Mendez", "2025-05-10", 0.38],
        ["Website Redesign", "Development", "Backend API migration", "Carlos Mendez", "2025-05-20", 0.22],
        ["Website Redesign", "Testing", "Cross-browser QA testing", "Nina Petrov", "2025-06-05", 0.10],
        ["Mobile App v3", "Planning", "Feature prioritization matrix", "David Kim", "2025-03-22", 0.88],
        ["Mobile App v3", "Planning", "Sprint roadmap creation", "David Kim", "2025-03-28", 0.75],
        ["Mobile App v3", "Design", "UI mockup iterations", "Priya Sharma", "2025-04-08", 0.60],
        ["Mobile App v3", "Design", "Accessibility audit checklist", "Priya Sharma", "2025-04-18", 0.48],
        ["Mobile App v3", "Development", "Push notification service", "James O'Brien", "2025-05-05", 0.35],
        ["Mobile App v3", "Development", "Offline sync module", "James O'Brien", "2025-05-25", 0.20],
        ["Mobile App v3", "Testing", "Performance benchmark suite", "Sofia Rossi", "2025-06-10", 0.08],
        ["Data Warehouse", "Planning", "Schema design review", "Elena Volkov", "2025-03-12", 0.97],
        ["Data Warehouse", "Planning", "ETL pipeline specification", "Elena Volkov", "2025-03-30", 0.82],
        ["Data Warehouse", "Design", "Dimensional model blueprint", "Hassan Ali", "2025-04-14", 0.70],
        ["Data Warehouse", "Design", "Security compliance mapping", "Hassan Ali", "2025-04-22", 0.53],
        ["Data Warehouse", "Development", "Ingestion pipeline build", "Rachel Torres", "2025-05-08", 0.45],
        ["Data Warehouse", "Development", "Dashboard connector setup", "Rachel Torres", "2025-05-18", 0.28],
        ["Data Warehouse", "Testing", "Data integrity validation", "Michael Chang", "2025-06-02", 0.18],
        ["CRM Integration", "Planning", "Integration scope document", "Laura Fischer", "2025-03-16", 0.90],
        ["CRM Integration", "Design", "API contract definition", "Yuki Tanaka", "2025-04-20", 0.65],
        ["CRM Integration", "Development", "Salesforce adapter module", "Omar Benali", "2025-05-12", 0.40],
        ["CRM Integration", "Development", "Contact sync automation", "Omar Benali", "2025-05-22", 0.12],
        ["CRM Integration", "Testing", "End-to-end workflow test", "Anna Lindgren", "2025-06-08", 0.05],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 6:  # Completion % column
                cell.number_format = '0%'
                cell.alignment = Alignment(horizontal="center")
            elif c == 5:  # Target date column
                cell.alignment = Alignment(horizontal="center")

    # No conditional formatting in initial state
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
