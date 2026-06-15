"""
Reward Script: Pivot table from survey data counting responses by age group and satisfaction level
Task ID: calc_pivot_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) - PivotTable sheet exists
  Component 2 (0.25) - Correct structure: AgeGroup rows, SatisfactionLevel columns, count header
  Component 3 (0.30) - Specific cell values match ground truth (18-24/Satisfied=22, 35-44/Neutral=18)
  Component 4 (0.20) - Grand total row/column present and grand total = 400
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_033'

# Expected age groups and satisfaction levels
EXPECTED_AGE_GROUPS = {'18-24', '25-34', '35-44', '45-54', '55+'}
EXPECTED_SAT_LEVELS = {'Very Unsatisfied', 'Unsatisfied', 'Neutral', 'Satisfied', 'Very Satisfied'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotTable sheet exists (0.25 points)
    # This sheet does NOT exist in initial_env, only in golden_env
    try:
        pivot_ws = None
        for sn in wb.sheetnames:
            if 'pivot' in sn.lower():
                pivot_ws = wb[sn]
                break
        if pivot_ws is not None:
            print(f"PASS: Component 1 - Pivot table sheet found: '{pivot_ws.title}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 - No sheet with 'pivot' in the name found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 2: Correct structure - AgeGroup as rows, SatisfactionLevel as columns (0.25 points)
    # Check that the header row contains satisfaction levels and the row labels contain age groups
    try:
        structure_score = 0.0

        # Read header row (row 1) for column labels - should contain satisfaction levels
        header_values = set()
        for col in range(2, pivot_ws.max_column + 1):
            val = pivot_ws.cell(row=1, column=col).value
            if val is not None:
                header_values.add(str(val).strip())

        # Check satisfaction levels are in the header
        sat_in_header = EXPECTED_SAT_LEVELS.intersection(header_values)
        if len(sat_in_header) >= 4:
            structure_score += 0.125
            print(f"  Structure check: {len(sat_in_header)}/5 satisfaction levels found in header")
        else:
            print(f"  Structure check: Only {len(sat_in_header)}/5 satisfaction levels in header: {sat_in_header}")

        # Read row labels (column A) for age groups
        row_labels = set()
        for row in range(2, pivot_ws.max_row + 1):
            val = pivot_ws.cell(row=row, column=1).value
            if val is not None:
                row_labels.add(str(val).strip())

        age_in_rows = EXPECTED_AGE_GROUPS.intersection(row_labels)
        if len(age_in_rows) >= 4:
            structure_score += 0.125
            print(f"  Structure check: {len(age_in_rows)}/5 age groups found in rows")
        else:
            print(f"  Structure check: Only {len(age_in_rows)}/5 age groups in rows: {age_in_rows}")

        if structure_score > 0:
            print(f"PASS: Component 2 - Pivot table structure verified ({structure_score} pts)")
            total_score += structure_score
        else:
            print(f"FAIL: Component 2 - Pivot table structure incorrect")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Helper: build a lookup dict from the pivot table for (age_group, sat_level) -> value
    def build_pivot_lookup(ws):
        """Build a dict mapping (row_label, col_label) -> cell value."""
        lookup = {}
        # Get column headers
        col_headers = {}
        for col in range(2, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            if val is not None:
                col_headers[col] = str(val).strip()
        # Get row data
        for row in range(2, ws.max_row + 1):
            row_label = ws.cell(row=row, column=1).value
            if row_label is not None:
                row_label = str(row_label).strip()
                for col, col_label in col_headers.items():
                    cell_val = ws.cell(row=row, column=col).value
                    lookup[(row_label, col_label)] = cell_val
        return lookup

    # Component 3: Specific ground truth values (0.30 points)
    # 18-24/Satisfied=22 and 35-44/Neutral=18
    try:
        lookup = build_pivot_lookup(pivot_ws)
        gt_checks = [
            ('18-24', 'Satisfied', 22, 0.15),
            ('35-44', 'Neutral', 18, 0.15),
        ]
        for age, sat, expected, pts in gt_checks:
            actual = lookup.get((age, sat))
            if actual is not None and int(actual) == expected:
                print(f"PASS: Component 3 - {age}/{sat} = {actual} (expected {expected}) ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 3 - {age}/{sat} = {actual} (expected {expected})")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Grand total = 400 (0.20 points)
    # Check for a grand total value of 400 in a row/column labeled "Grand Total" or similar
    try:
        comp4_score = 0.0
        for row in range(1, pivot_ws.max_row + 1):
            for col in range(1, pivot_ws.max_column + 1):
                val = pivot_ws.cell(row=row, column=col).value
                if val is not None and comp4_score == 0.0:
                    try:
                        if int(val) == 400:
                            row_label = pivot_ws.cell(row=row, column=1).value
                            col_label = pivot_ws.cell(row=1, column=col).value
                            if (row_label and 'total' in str(row_label).lower()) or \
                               (col_label and 'total' in str(col_label).lower()):
                                comp4_score = 0.20
                                print(f"PASS: Component 4 - Grand total 400 found at row={row}, col={col} (0.20 pts)")
                    except (ValueError, TypeError):
                        pass

        if comp4_score > 0:
            total_score += comp4_score
        else:
            print(f"FAIL: Component 4 - Grand total of 400 not found in a 'Total' row/column")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
