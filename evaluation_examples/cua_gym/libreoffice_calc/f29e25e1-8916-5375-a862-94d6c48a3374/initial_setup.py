"""
Initial Setup: Survey data for pivot table task
Task ID: calc_pivot_033
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    age_groups = ['18-24', '25-34', '35-44', '45-54', '55+']
    satisfaction_levels = ['Very Unsatisfied', 'Unsatisfied', 'Neutral', 'Satisfied', 'Very Satisfied']

    # We need specific counts: 18-24/Satisfied=22, 35-44/Neutral=18, Grand total=400
    # Build a distribution matrix (age_group x satisfaction) that sums to 400
    # 5 age groups x 5 satisfaction = 25 cells, average ~16 each
    count_matrix = {
        '18-24':  {'Very Unsatisfied': 14, 'Unsatisfied': 16, 'Neutral': 19, 'Satisfied': 22, 'Very Satisfied': 15},
        '25-34':  {'Very Unsatisfied': 12, 'Unsatisfied': 18, 'Neutral': 21, 'Satisfied': 23, 'Very Satisfied': 16},
        '35-44':  {'Very Unsatisfied': 15, 'Unsatisfied': 16, 'Neutral': 18, 'Satisfied': 17, 'Very Satisfied': 14},
        '45-54':  {'Very Unsatisfied': 16, 'Unsatisfied': 14, 'Neutral': 15, 'Satisfied': 19, 'Very Satisfied': 12},
        '55+':    {'Very Unsatisfied': 11, 'Unsatisfied': 13, 'Neutral': 16, 'Satisfied': 16, 'Very Satisfied': 12},
    }

    # Verify total = 400
    total = sum(v for ag in count_matrix.values() for v in ag.values())
    assert total == 400, f"Total is {total}, expected 400"
    # Verify specific counts
    assert count_matrix['18-24']['Satisfied'] == 22
    assert count_matrix['35-44']['Neutral'] == 18

    # Generate 400 rows of data based on the distribution
    rows = []
    resp_id = 1
    for ag in age_groups:
        for sat in satisfaction_levels:
            n = count_matrix[ag][sat]
            for _ in range(n):
                # Score: correlate loosely with satisfaction
                base_scores = {
                    'Very Unsatisfied': (1, 3),
                    'Unsatisfied': (2, 5),
                    'Neutral': (4, 6),
                    'Satisfied': (6, 8),
                    'Very Satisfied': (8, 10),
                }
                lo, hi = base_scores[sat]
                score = random.randint(lo, hi)

                # Comments
                comment_templates = {
                    'Very Unsatisfied': [
                        "Very disappointed with the service quality.",
                        "Would not recommend to anyone.",
                        "Terrible experience overall.",
                        "Needs major improvements in every area.",
                        "Extremely poor customer support.",
                    ],
                    'Unsatisfied': [
                        "Below expectations unfortunately.",
                        "Several issues need to be addressed.",
                        "Could be much better.",
                        "Not worth the price paid.",
                        "Disappointing in multiple areas.",
                    ],
                    'Neutral': [
                        "Average experience, nothing special.",
                        "Some good points, some bad.",
                        "Met basic expectations only.",
                        "Neither impressed nor disappointed.",
                        "Room for improvement but acceptable.",
                    ],
                    'Satisfied': [
                        "Good overall experience.",
                        "Happy with the service provided.",
                        "Would consider returning again.",
                        "Most aspects met my expectations.",
                        "Solid experience, minor issues only.",
                    ],
                    'Very Satisfied': [
                        "Exceptional service, highly recommend!",
                        "Exceeded all my expectations.",
                        "Outstanding quality and attention.",
                        "Best experience I have had.",
                        "Will definitely come back again!",
                    ],
                }
                comment = random.choice(comment_templates[sat])
                rows.append((resp_id, ag, sat, score, comment))
                resp_id += 1

    # Shuffle to make it realistic (not grouped by age/satisfaction)
    random.shuffle(rows)
    # Re-assign sequential RespIDs after shuffle
    rows = [(i + 1, r[1], r[2], r[3], r[4]) for i, r in enumerate(rows)]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SurveyResults'

    # Headers
    headers = ['RespID', 'AgeGroup', 'SatisfactionLevel', 'Score', 'Comments']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Data rows
    for r, row_data in enumerate(rows, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 45

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
