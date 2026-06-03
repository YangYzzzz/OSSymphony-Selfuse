"""
Reward Script: Add EMNLP 2019 and 2020 best paper awards to LibreOffice Calc
Task ID: osworld_multi_apps_acl_awards_calc_006
Domain: libreoffice_calc
Scoring:
  Component 1: Two data rows added (file has 3 rows total)       — 0.2 pts
  Component 2: 2019 row has correct Year and paper title         — 0.3 pts
  Component 3: 2019 row has correct authors and Affiliation Type — 0.1 pts
  Component 4: 2020 row has correct Year and paper title         — 0.3 pts
  Component 5: 2020 row has correct authors and Affiliation Type — 0.1 pts
  Total: 1.0
"""

import os
import shutil
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_acl_awards_calc_006'

# Ground truth values from golden_env
GOLDEN_2019_TITLE_KEYWORDS = ['specializing', 'word', 'embeddings', 'information bottleneck']
GOLDEN_2019_AUTHORS = ['xiang', 'lisa', 'li', 'jason', 'eisner']
GOLDEN_2019_YEAR = 2019

GOLDEN_2020_TITLE_KEYWORDS = ['digital', 'voicing', 'silent', 'speech']
GOLDEN_2020_AUTHORS = ['gaddy', 'klein']
GOLDEN_2020_YEAR = 2020

VALID_AFFILIATION_TYPES = {'academic', 'industry', 'mixed'}


def normalize(text):
    """Normalize text for comparison: lowercase and strip whitespace."""
    if text is None:
        return ''
    return str(text).strip().lower()


def load_ods_as_openpyxl(file_path):
    """
    Load an ODS file using openpyxl by copying to .xlsx extension.
    ODS files saved by LibreOffice in xlsx-compat format can be read this way.
    """
    tmp_path = '/tmp/emnlp_awards_reward_check.xlsx'
    shutil.copy(file_path, tmp_path)
    wb = openpyxl.load_workbook(tmp_path)
    return wb


def find_row_for_year(ws, year):
    """Find the first row (after header) that has the given year in column A."""
    for row_idx in range(2, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=1).value
        try:
            if int(val) == year:
                return row_idx
        except (TypeError, ValueError):
            pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = load_ods_as_openpyxl(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get active sheet
    try:
        ws = wb.active
        print(f"INFO: Active sheet name: '{ws.title}'")
        print(f"INFO: max_row={ws.max_row}, max_col={ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Print all rows for debugging
    print("\nINFO: All rows in file:")
    for r_idx in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r_idx, column=c).value for c in range(1, ws.max_column + 1)]
        print(f"  Row {r_idx}: {row_vals}")

    # Component 1: Two data rows added (file has 3 rows total: header + 2019 + 2020) (0.2 pts)
    try:
        # Count non-empty data rows (excluding header row 1)
        data_rows = 0
        for r_idx in range(2, ws.max_row + 1):
            # Check if row has any non-None values
            row_has_data = any(
                ws.cell(row=r_idx, column=c).value is not None
                for c in range(1, ws.max_column + 1)
            )
            if row_has_data:
                data_rows += 1

        if data_rows >= 2:
            print(f"PASS: Component 1 — {data_rows} data rows found (expected >= 2) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — only {data_rows} data row(s) found, expected 2 (one for 2019, one for 2020)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Find rows for 2019 and 2020
    row_2019 = find_row_for_year(ws, 2019)
    row_2020 = find_row_for_year(ws, 2020)
    print(f"\nINFO: Row index for year 2019: {row_2019}")
    print(f"INFO: Row index for year 2020: {row_2020}")

    # Component 2: 2019 row has correct Year and paper title (0.3 pts)
    try:
        if row_2019 is not None:
            title_val = normalize(ws.cell(row=row_2019, column=2).value)
            # Check title contains at least 2 of the key keywords (flexible matching)
            matching_keywords = sum(1 for kw in GOLDEN_2019_TITLE_KEYWORDS if kw in title_val)
            if matching_keywords >= 2:
                print(f"PASS: Component 2 — 2019 row found with matching title "
                      f"(keywords matched: {matching_keywords}/{len(GOLDEN_2019_TITLE_KEYWORDS)}) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — 2019 row found but title mismatch: '{title_val}' "
                      f"(expected keywords: {GOLDEN_2019_TITLE_KEYWORDS}, matched: {matching_keywords})")
        else:
            print(f"FAIL: Component 2 — no row with year 2019 found in the file")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 2019 row has correct authors and Affiliation Type (0.1 pts)
    try:
        if row_2019 is not None:
            authors_val = normalize(ws.cell(row=row_2019, column=3).value)
            affiliation_val = normalize(ws.cell(row=row_2019, column=4).value)

            # Check that at least some expected author names are present
            matching_authors = sum(1 for name in GOLDEN_2019_AUTHORS if name in authors_val)
            affiliation_valid = affiliation_val in VALID_AFFILIATION_TYPES

            if matching_authors >= 2 and affiliation_valid:
                print(f"PASS: Component 3 — 2019 row has valid authors ({authors_val}) "
                      f"and valid Affiliation Type ({affiliation_val}) (0.1 pts)")
                total_score += 0.1
            else:
                issues = []
                if matching_authors < 2:
                    issues.append(f"authors mismatch: '{authors_val}' (matched {matching_authors} expected names)")
                if not affiliation_valid:
                    issues.append(f"affiliation type invalid: '{affiliation_val}' "
                                  f"(expected one of {VALID_AFFILIATION_TYPES})")
                print(f"FAIL: Component 3 — 2019 row issues: {'; '.join(issues)}")
        else:
            print(f"FAIL: Component 3 — no row with year 2019 found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: 2020 row has correct Year and paper title (0.3 pts)
    try:
        if row_2020 is not None:
            title_val = normalize(ws.cell(row=row_2020, column=2).value)
            # Check title contains at least 2 of the key keywords (flexible matching)
            matching_keywords = sum(1 for kw in GOLDEN_2020_TITLE_KEYWORDS if kw in title_val)
            if matching_keywords >= 2:
                print(f"PASS: Component 4 — 2020 row found with matching title "
                      f"(keywords matched: {matching_keywords}/{len(GOLDEN_2020_TITLE_KEYWORDS)}) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 4 — 2020 row found but title mismatch: '{title_val}' "
                      f"(expected keywords: {GOLDEN_2020_TITLE_KEYWORDS}, matched: {matching_keywords})")
        else:
            print(f"FAIL: Component 4 — no row with year 2020 found in the file")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: 2020 row has correct authors and Affiliation Type (0.1 pts)
    try:
        if row_2020 is not None:
            authors_val = normalize(ws.cell(row=row_2020, column=3).value)
            affiliation_val = normalize(ws.cell(row=row_2020, column=4).value)

            # Check that at least some expected author names are present
            matching_authors = sum(1 for name in GOLDEN_2020_AUTHORS if name in authors_val)
            affiliation_valid = affiliation_val in VALID_AFFILIATION_TYPES

            if matching_authors >= 1 and affiliation_valid:
                print(f"PASS: Component 5 — 2020 row has valid authors ({authors_val}) "
                      f"and valid Affiliation Type ({affiliation_val}) (0.1 pts)")
                total_score += 0.1
            else:
                issues = []
                if matching_authors < 1:
                    issues.append(f"authors mismatch: '{authors_val}' (matched {matching_authors} expected names)")
                if not affiliation_valid:
                    issues.append(f"affiliation type invalid: '{affiliation_val}' "
                                  f"(expected one of {VALID_AFFILIATION_TYPES})")
                print(f"FAIL: Component 5 — 2020 row issues: {'; '.join(issues)}")
        else:
            print(f"FAIL: Component 5 — no row with year 2020 found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM environment
file_path = f'{WORKDIR}/emnlp_awards.ods'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
