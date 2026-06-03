"""
Initial Setup: ACL Awards EMNLP Best Papers - LibreOffice Calc
Task ID: osworld_multi_apps_acl_awards_calc_006
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Description:
  Creates emnlp_awards.ods with only header row (Year, Paper Title, Authors,
  Affiliation Type). The agent must look up EMNLP 2019 and 2020 best paper
  winners in Chrome and add the data rows.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_acl_awards_calc_006'
# Save as .xlsx (openpyxl format); LibreOffice Calc will open it as emnlp_awards.xlsx
# The task refers to "emnlp_awards.ods" but openpyxl saves xlsx format.
# We use .xlsx extension to ensure valid format detection by both LibreOffice and reward-gen.
OUTPUT = f'{WORKDIR}/emnlp_awards.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: EMNLP Awards ---
    ws = wb.active
    ws.title = 'EMNLP Awards'

    # Column headers
    headers = ['Year', 'Paper Title', 'Authors', 'Affiliation Type']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=12, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Set column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20

    # Row height for header
    ws.row_dimensions[1].height = 22

    # NO data rows — the agent must look up and fill in EMNLP 2019 and 2020

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open Chrome first, then LibreOffice Calc
    # Chrome for browsing aclweb.org/aclwiki/Best_paper_awards
    launch_gui('google-chrome', delay_sec=3.0)

    # Open LibreOffice Calc with the awards file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
