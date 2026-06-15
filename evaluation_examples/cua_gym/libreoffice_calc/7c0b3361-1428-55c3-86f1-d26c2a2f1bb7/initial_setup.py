"""
Initial Setup: Create a multi-sheet workbook for tab-delimited export task
Task ID: calc_gsi_057
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws1 = wb.active
    ws1.title = 'Inventory'

    headers1 = ['Item Code', 'Product Name', 'Category', 'Unit Price', 'Qty In Stock', 'Reorder Level', 'Last Updated']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    inventory_data = [
        ['INV-1001', 'Precision Ball Bearing 6205', 'Mechanical Parts', 12.50, 340, 100, '2025-11-02'],
        ['INV-1002', 'Stainless Steel Flange DN50', 'Pipe Fittings', 28.75, 185, 50, '2025-10-28'],
        ['INV-1003', 'Hydraulic Cylinder 40mm Bore', 'Hydraulics', 145.00, 42, 20, '2025-11-05'],
        ['INV-1004', 'Industrial V-Belt B68', 'Drive Components', 8.90, 520, 200, '2025-09-15'],
        ['INV-1005', 'Pneumatic Solenoid Valve 24V', 'Pneumatics', 67.30, 78, 30, '2025-11-01'],
        ['INV-1006', 'Carbon Steel Pipe 2 inch', 'Pipe Fittings', 34.20, 290, 100, '2025-10-20'],
        ['INV-1007', 'Linear Guide Rail 500mm', 'Motion Control', 89.50, 55, 25, '2025-11-03'],
        ['INV-1008', 'Thermoplastic Coupling GR28', 'Drive Components', 22.15, 165, 60, '2025-10-12'],
        ['INV-1009', 'Pressure Gauge 0-10 Bar', 'Instrumentation', 18.40, 210, 80, '2025-11-04'],
        ['INV-1010', 'Gear Motor 0.75kW 1400rpm', 'Motors', 312.00, 18, 10, '2025-10-30'],
        ['INV-1011', 'Roller Chain 10B-1 5m', 'Drive Components', 45.60, 92, 40, '2025-09-28'],
        ['INV-1012', 'Proximity Sensor NPN 8mm', 'Sensors', 24.80, 148, 50, '2025-11-06'],
        ['INV-1013', 'Shaft Seal TC 35x52x7', 'Seals & Gaskets', 3.75, 780, 300, '2025-10-15'],
        ['INV-1014', 'Aluminum Extrusion 40x40 2m', 'Structural', 19.90, 125, 50, '2025-10-22'],
    ]
    for r, row_data in enumerate(inventory_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 32
    ws1.column_dimensions['C'].width = 18
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 14
    ws1.column_dimensions['F'].width = 14
    ws1.column_dimensions['G'].width = 14

    # --- Sheet 2: Suppliers ---
    ws2 = wb.create_sheet('Suppliers')

    headers2 = ['Supplier ID', 'Company Name', 'Contact Person', 'Phone', 'Email', 'City', 'Rating']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="FF548235", end_color="FF548235", fill_type="solid")
        cell.alignment = header_align

    supplier_data = [
        ['SUP-201', 'Pacific Industrial Supply Co.', 'Robert Tanaka', '+1-503-555-0142', 'rtanaka@pacind.com', 'Portland', 4.5],
        ['SUP-202', 'Eurotech Components GmbH', 'Annika Braun', '+49-89-555-0188', 'a.braun@eurotech.de', 'Munich', 4.8],
        ['SUP-203', 'Shanghai Mechanical Parts Ltd.', 'Wei Chen', '+86-21-555-0276', 'wchen@shmp.cn', 'Shanghai', 4.2],
        ['SUP-204', 'Great Lakes Bearings Inc.', 'Jennifer Walsh', '+1-312-555-0359', 'jwalsh@glbearings.com', 'Chicago', 4.6],
        ['SUP-205', 'Nordic Hydraulics AB', 'Lars Eriksson', '+46-31-555-0431', 'l.eriksson@nordhyd.se', 'Gothenburg', 4.7],
        ['SUP-206', 'Atlas Automation Pvt Ltd', 'Priya Sharma', '+91-80-555-0522', 'psharma@atlasauto.in', 'Bangalore', 4.1],
        ['SUP-207', 'Precision Parts Direct LLC', 'Michael Torres', '+1-214-555-0618', 'mtorres@ppd.com', 'Dallas', 4.4],
        ['SUP-208', 'Kanto Steel Products KK', 'Yuki Nakamura', '+81-3-555-0743', 'y.nakamura@kantosteel.jp', 'Tokyo', 4.9],
        ['SUP-209', 'Midlands Engineering Supplies', 'James Crawford', '+44-121-555-0867', 'j.crawford@midseng.co.uk', 'Birmingham', 4.3],
        ['SUP-210', 'Sao Paulo Valve & Fitting', 'Carlos Ferreira', '+55-11-555-0912', 'cferreira@spvf.com.br', 'Sao Paulo', 4.0],
    ]
    for r, row_data in enumerate(supplier_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 32
    ws2.column_dimensions['C'].width = 20
    ws2.column_dimensions['D'].width = 18
    ws2.column_dimensions['E'].width = 28
    ws2.column_dimensions['F'].width = 14
    ws2.column_dimensions['G'].width = 10

    # --- Sheet 3: Orders ---
    ws3 = wb.create_sheet('Orders')

    headers3 = ['Order No', 'Order Date', 'Supplier ID', 'Item Code', 'Qty Ordered', 'Unit Cost', 'Total Cost', 'Status']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = PatternFill(start_color="FFBF8F00", end_color="FFBF8F00", fill_type="solid")
        cell.alignment = header_align

    order_data = [
        ['PO-3001', '2025-10-01', 'SUP-201', 'INV-1004', 500, 8.50, 4250.00, 'Delivered'],
        ['PO-3002', '2025-10-03', 'SUP-204', 'INV-1001', 200, 11.80, 2360.00, 'Delivered'],
        ['PO-3003', '2025-10-07', 'SUP-205', 'INV-1003', 30, 138.00, 4140.00, 'In Transit'],
        ['PO-3004', '2025-10-10', 'SUP-202', 'INV-1007', 40, 85.20, 3408.00, 'Delivered'],
        ['PO-3005', '2025-10-14', 'SUP-203', 'INV-1013', 1000, 3.20, 3200.00, 'Delivered'],
        ['PO-3006', '2025-10-18', 'SUP-208', 'INV-1006', 150, 32.50, 4875.00, 'In Transit'],
        ['PO-3007', '2025-10-22', 'SUP-206', 'INV-1005', 50, 64.00, 3200.00, 'Pending'],
        ['PO-3008', '2025-10-25', 'SUP-207', 'INV-1012', 100, 23.50, 2350.00, 'Delivered'],
        ['PO-3009', '2025-10-28', 'SUP-209', 'INV-1008', 80, 21.00, 1680.00, 'In Transit'],
        ['PO-3010', '2025-11-01', 'SUP-210', 'INV-1009', 120, 17.50, 2100.00, 'Pending'],
        ['PO-3011', '2025-11-03', 'SUP-201', 'INV-1011', 60, 43.20, 2592.00, 'Pending'],
        ['PO-3012', '2025-11-05', 'SUP-204', 'INV-1014', 75, 18.90, 1417.50, 'Processing'],
    ]
    for r, row_data in enumerate(order_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 12
    ws3.column_dimensions['B'].width = 14
    ws3.column_dimensions['C'].width = 14
    ws3.column_dimensions['D'].width = 14
    ws3.column_dimensions['E'].width = 14
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 12
    ws3.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
