"""
Reward Script: Professional Invoice in LibreOffice Calc
Task ID: calc_grs_001
Domain: libreoffice_calc
Scoring:
  Component 1: Merged header A1:F1 with company name (0.15)
  Component 2: Invoice detail rows 3-5 (0.10)
  Component 3: Items table header row 7 bold with correct columns (0.15)
  Component 4: At least 5 line items with quantities and prices (0.20)
  Component 5: Currency formatting on monetary columns (0.10)
  Component 6: SUM/Tax/Grand Total formulas in totals section (0.20)
  Component 7: Thick border around the invoice (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_001'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Merged header A1:F1 with "Acme Consulting LLC" (0.15 points)
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in str(r) for r in ws.merged_cells.ranges)
        header_val = ws['A1'].value
        if has_merge and header_val and 'Acme Consulting' in str(header_val):
            print(f"PASS: Component 1 — Merged header found: '{header_val}' in ranges {merged_ranges} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Expected merged A1:F1 with 'Acme Consulting LLC'. merge={has_merge}, val={header_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Invoice detail rows 3-5 (Invoice #, Date, Due Date) (0.10 points)
    try:
        details_found = 0
        # Check for Invoice # label
        for row in range(3, 6):
            val = ws.cell(row=row, column=1).value
            if val and 'invoice' in str(val).lower():
                details_found += 1
            if val and 'date' in str(val).lower():
                details_found += 1
            if val and 'due' in str(val).lower():
                details_found += 1

        # Need at least invoice number, date, and due date labels
        if details_found >= 3:
            print(f"PASS: Component 2 — Invoice details found ({details_found} labels) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Expected Invoice #, Date, Due Date in rows 3-5. Found {details_found} labels.")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Items table header row 7 with correct column names, all bold (0.15 points)
    try:
        expected_headers = ['item description', 'quantity', 'unit price', 'discount %', 'discount amount', 'total']
        headers_ok = 0
        bold_ok = 0
        for col in range(1, 7):
            cell = ws.cell(row=7, column=col)
            val = str(cell.value).strip().lower() if cell.value else ''
            # Check if the header roughly matches (contains key words)
            if col <= len(expected_headers) and expected_headers[col - 1] in val:
                headers_ok += 1
            elif cell.value is not None:
                # Partial match: accept if any expected header keyword is present
                for eh in expected_headers:
                    kw = eh.split()[0]  # first word
                    if kw in val:
                        headers_ok += 1
                        break
            if not isinstance(cell, MergedCell) and cell.font.bold:
                bold_ok += 1

        if headers_ok >= 5 and bold_ok >= 5:
            print(f"PASS: Component 3 — Header row 7 has {headers_ok}/6 correct headers, {bold_ok}/6 bold (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Header row: {headers_ok}/6 headers matched, {bold_ok}/6 bold")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: At least 5 line items with quantities and prices (0.20 points)
    try:
        line_items = 0
        # Scan rows 8-20 for line items (description in col A, quantity in col B, price in col C)
        for row in range(8, 21):
            desc = ws.cell(row=row, column=1).value
            qty = ws.cell(row=row, column=2).value
            price = ws.cell(row=row, column=3).value
            if desc is not None and qty is not None and price is not None:
                try:
                    float(qty)
                    float(price)
                    line_items += 1
                except (ValueError, TypeError):
                    pass

        if line_items >= 5:
            print(f"PASS: Component 4 — Found {line_items} line items (>= 5 required) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Found only {line_items} line items, need >= 5")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Currency formatting on monetary columns (C, E, F in item rows) (0.10 points)
    try:
        currency_cells = 0
        total_checked = 0
        # Check unit price (C), discount amount (E), total (F) for rows with line items
        for row in range(8, 21):
            if ws.cell(row=row, column=1).value is None:
                continue
            for col in [3, 5, 6]:  # C=Unit Price, E=Discount Amount, F=Total
                total_checked += 1
                nf = ws.cell(row=row, column=col).number_format
                if nf and ('$' in str(nf) or '#,##0' in str(nf)):
                    currency_cells += 1

        if total_checked > 0 and currency_cells / total_checked >= 0.7:
            print(f"PASS: Component 5 — Currency formatting: {currency_cells}/{total_checked} cells formatted (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Currency formatting: only {currency_cells}/{total_checked} cells have currency format")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: SUM formula for Subtotal, Tax (8.5%), Grand Total (0.20 points)
    try:
        formulas_found = 0

        # Look for Subtotal with SUM formula in rows 13-20
        for row in range(13, 21):
            label = ws.cell(row=row, column=5).value
            formula = ws.cell(row=row, column=6).value
            if label and 'subtotal' in str(label).lower():
                if formula and 'SUM' in str(formula).upper():
                    formulas_found += 1
                    print(f"  Found Subtotal SUM at row {row}: {formula}")
                    break
            # Also check col 4 for label
            label4 = ws.cell(row=row, column=4).value
            if label4 and 'subtotal' in str(label4).lower():
                formula5 = ws.cell(row=row, column=5).value
                formula6 = ws.cell(row=row, column=6).value
                f = formula5 or formula6
                if f and 'SUM' in str(f).upper():
                    formulas_found += 1
                    print(f"  Found Subtotal SUM at row {row}: {f}")
                    break

        # Look for Tax formula with 0.085 or 8.5
        for row in range(13, 21):
            for col in [4, 5]:
                label = ws.cell(row=row, column=col).value
                if label and 'tax' in str(label).lower():
                    formula = ws.cell(row=row, column=6).value or ws.cell(row=row, column=col+1).value
                    if formula and ('0.085' in str(formula) or '8.5' in str(formula)):
                        formulas_found += 1
                        print(f"  Found Tax formula at row {row}: {formula}")
                    break

        # Look for Grand Total formula
        for row in range(13, 21):
            for col in [4, 5]:
                label = ws.cell(row=row, column=col).value
                if label and 'grand' in str(label).lower() and 'total' in str(label).lower():
                    formula = ws.cell(row=row, column=6).value or ws.cell(row=row, column=col+1).value
                    if formula and isinstance(formula, str) and '=' in formula:
                        formulas_found += 1
                        print(f"  Found Grand Total formula at row {row}: {formula}")
                    break

        sub_score = formulas_found * (0.20 / 3.0)
        if formulas_found >= 3:
            print(f"PASS: Component 6 — All 3 formulas found (Subtotal SUM, Tax 8.5%, Grand Total) (0.20 pts)")
            total_score += 0.20
        elif formulas_found > 0:
            print(f"PARTIAL: Component 6 — {formulas_found}/3 formulas found ({sub_score:.2f} pts)")
            total_score += round(sub_score, 2)
        else:
            print(f"FAIL: Component 6 — No Subtotal/Tax/Grand Total formulas found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Thick border around the invoice (0.10 points)
    try:
        thick_borders_found = 0
        # Check for thick borders on outer edges
        # Top-left corner (A1)
        if ws['A1'].border.left.style == 'thick' and ws['A1'].border.top.style == 'thick':
            thick_borders_found += 1
        # Top-right corner (F1)
        f1 = ws['F1']
        if not isinstance(f1, MergedCell):
            if f1.border.right.style == 'thick' and f1.border.top.style == 'thick':
                thick_borders_found += 1
        else:
            thick_borders_found += 1  # merged cell, check A1 right side handled below

        # Check left edge on a middle row
        last_data_row = None
        for row in range(20, 6, -1):
            if ws.cell(row=row, column=6).value is not None or ws.cell(row=row, column=5).value is not None:
                last_data_row = row
                break

        if last_data_row:
            mid_row = (8 + last_data_row) // 2
            if ws.cell(row=mid_row, column=1).border.left.style == 'thick':
                thick_borders_found += 1
            # Bottom-right corner
            if ws.cell(row=last_data_row, column=6).border.bottom.style == 'thick':
                thick_borders_found += 1

        if thick_borders_found >= 3:
            print(f"PASS: Component 7 — Thick borders detected on {thick_borders_found} edges (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 — Only {thick_borders_found} thick border edges found (need >= 3)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
