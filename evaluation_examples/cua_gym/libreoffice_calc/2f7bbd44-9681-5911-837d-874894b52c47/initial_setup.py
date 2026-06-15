"""
Initial Setup: European date misinterpretation in LibreOffice Calc
Task ID: calc_tbl_020
Domain: libreoffice_calc

D2:D30 contain text dates in DD/MM/YYYY format that were entered as text strings.
Calc has not converted them — they sit as plain text showing the DD/MM/YYYY pattern.
The agent must fix these to proper date values reflecting European interpretation.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from datetime import datetime

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Orders ---
    ws = wb.active
    ws.title = 'Orders'

    # Headers
    headers = ['Order ID', 'Customer', 'Product', 'Order Date', 'Amount', 'Status']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Data — 29 rows of orders (D2:D30)
    # Dates are TEXT strings in DD/MM/YYYY format
    # Some have day > 12 (unambiguous), some have day <= 12 (ambiguous, the tricky ones)
    orders = [
        ['ORD-1001', 'Sarah Chen',         'Ergonomic Office Chair',       '05/02/2024', 349.99, 'Delivered'],
        ['ORD-1002', 'Marcus Johnson',      'Wireless Noise-Cancel Headphones', '12/03/2024', 229.50, 'Shipped'],
        ['ORD-1003', 'Elena Rodriguez',     'Standing Desk Converter',      '23/01/2024', 415.00, 'Delivered'],
        ['ORD-1004', 'James O\'Brien',      'Mechanical Keyboard RGB',      '07/04/2024', 159.95, 'Processing'],
        ['ORD-1005', 'Priya Patel',         'Ultra-Wide Monitor 34"',       '15/06/2024', 699.00, 'Delivered'],
        ['ORD-1006', 'Thomas Andersen',     'Laptop Cooling Pad',           '01/08/2024', 45.99,  'Shipped'],
        ['ORD-1007', 'Yuki Tanaka',         'USB-C Docking Station',        '28/02/2024', 189.00, 'Delivered'],
        ['ORD-1008', 'Amara Okonkwo',       'Webcam 4K Ultra HD',           '09/11/2023', 134.50, 'Returned'],
        ['ORD-1009', 'David Kim',           'Ergonomic Mouse Vertical',     '11/05/2024', 69.99,  'Delivered'],
        ['ORD-1010', 'Isabelle Moreau',     'Cable Management Kit',         '03/07/2024', 32.99,  'Shipped'],
        ['ORD-1011', 'Roberto Vega',        'Monitor Arm Dual Mount',       '18/09/2023', 124.00, 'Delivered'],
        ['ORD-1012', 'Fatima Al-Hassan',    'Privacy Screen Filter 27"',    '06/01/2024', 54.99,  'Delivered'],
        ['ORD-1013', 'Nathan Brooks',       'Surge Protector 12-Outlet',    '25/11/2023', 39.95,  'Shipped'],
        ['ORD-1014', 'Ling Wei',            'Desk Lamp LED Adjustable',     '02/10/2024', 78.50,  'Processing'],
        ['ORD-1015', 'Sofia Martinez',      'Wireless Trackpad Multi-Touch','14/03/2024', 99.00,  'Delivered'],
        ['ORD-1016', 'Oliver Schwartz',     'Laptop Stand Aluminum',        '08/06/2024', 64.99,  'Shipped'],
        ['ORD-1017', 'Aisha Mohammed',      'Portable SSD 2TB',            '30/04/2024', 189.99, 'Delivered'],
        ['ORD-1018', 'Liam Fitzgerald',     'Bluetooth Speaker Desktop',    '10/12/2023', 85.00,  'Delivered'],
        ['ORD-1019', 'Chen Xiaoming',       'Wrist Rest Gel Pad',          '21/07/2024', 24.50,  'Shipped'],
        ['ORD-1020', 'Maria Gonzalez',      'USB Hub 7-Port Powered',      '04/09/2023', 42.99,  'Delivered'],
        ['ORD-1021', 'Henrik Larsson',      'Anti-Fatigue Floor Mat',      '16/02/2024', 59.00,  'Returned'],
        ['ORD-1022', 'Rachel Thompson',     'HDMI Cable 4K 3m',           '27/05/2024', 18.99,  'Delivered'],
        ['ORD-1023', 'Ahmed Yousef',        'Desk Organizer Bamboo',       '09/08/2024', 36.50,  'Processing'],
        ['ORD-1024', 'Keiko Nakamura',      'Screen Cleaning Kit Pro',     '01/11/2023', 15.99,  'Delivered'],
        ['ORD-1025', 'Patrick Dubois',      'Ergonomic Footrest',          '13/04/2024', 49.95,  'Shipped'],
        ['ORD-1026', 'Grace Okafor',        'Wireless Presenter Remote',   '22/06/2024', 35.00,  'Delivered'],
        ['ORD-1027', 'Dmitri Petrov',       'Monitor Light Bar',           '07/01/2024', 62.50,  'Delivered'],
        ['ORD-1028', 'Isabella Rossi',      'Noise Machine White Noise',   '19/10/2023', 44.99,  'Shipped'],
        ['ORD-1029', 'William Chang',       'Ethernet Adapter USB-C',      '11/03/2024', 29.99,  'Delivered'],
    ]

    for r, row_data in enumerate(orders, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Order ID
        ws.cell(row=r, column=2, value=row_data[1])  # Customer
        ws.cell(row=r, column=3, value=row_data[2])  # Product
        # Date as TEXT string — this is the key: store as plain text, not date
        date_cell = ws.cell(row=r, column=4, value=row_data[3])
        date_cell.number_format = '@'  # Text format to ensure it stays as text
        ws.cell(row=r, column=5, value=row_data[4])  # Amount
        ws.cell(row=r, column=6, value=row_data[5])  # Status

    # Format amount column as currency
    for r in range(2, 31):
        ws.cell(row=r, column=5).number_format = '$#,##0.00'

    # Freeze header row
    ws.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    ws2['A1'] = 'Order Summary'
    ws2['A1'].font = Font(size=14, bold=True)
    ws2['A3'] = 'Total Orders'
    ws2['B3'] = 29
    ws2['A4'] = 'Date Range'
    ws2['B4'] = 'Oct 2023 - Oct 2024'
    ws2['A5'] = 'Status Breakdown'
    ws2['A6'] = 'Delivered'
    ws2['B6'] = 17
    ws2['A7'] = 'Shipped'
    ws2['B7'] = 8
    ws2['A8'] = 'Processing'
    ws2['B8'] = 3
    ws2['A9'] = 'Returned'
    ws2['B9'] = 2

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
