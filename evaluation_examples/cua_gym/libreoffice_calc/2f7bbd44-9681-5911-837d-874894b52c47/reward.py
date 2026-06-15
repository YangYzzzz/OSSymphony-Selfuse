"""
Reward Script: Fix European DD/MM/YYYY date strings to proper datetime values
Task ID: calc_tbl_020
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): All D2:D30 cells are datetime objects (not text strings)
  Component 2 (0.35): Dates correctly reflect DD/MM/YYYY European interpretation
  Component 3 (0.25): Number format is a date format (not text '@')
"""

import os
from datetime import datetime

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_020'

# Expected dates based on DD/MM/YYYY interpretation of the original text strings
# Original text -> correct European datetime(year, month, day)
EXPECTED_DATES = {
    2:  datetime(2024, 2, 5),    # '05/02/2024' -> Feb 5
    3:  datetime(2024, 3, 12),   # '12/03/2024' -> Mar 12
    4:  datetime(2024, 1, 23),   # '23/01/2024' -> Jan 23
    5:  datetime(2024, 4, 7),    # '07/04/2024' -> Apr 7
    6:  datetime(2024, 6, 15),   # '15/06/2024' -> Jun 15
    7:  datetime(2024, 8, 1),    # '01/08/2024' -> Aug 1
    8:  datetime(2024, 2, 28),   # '28/02/2024' -> Feb 28
    9:  datetime(2023, 11, 9),   # '09/11/2023' -> Nov 9
    10: datetime(2024, 5, 11),   # '11/05/2024' -> May 11
    11: datetime(2024, 7, 3),    # '03/07/2024' -> Jul 3
    12: datetime(2023, 9, 18),   # '18/09/2023' -> Sep 18
    13: datetime(2024, 1, 6),    # '06/01/2024' -> Jan 6
    14: datetime(2023, 11, 25),  # '25/11/2023' -> Nov 25
    15: datetime(2024, 10, 2),   # '02/10/2024' -> Oct 2
    16: datetime(2024, 3, 14),   # '14/03/2024' -> Mar 14
    17: datetime(2024, 6, 8),    # '08/06/2024' -> Jun 8
    18: datetime(2024, 4, 30),   # '30/04/2024' -> Apr 30
    19: datetime(2023, 12, 10),  # '10/12/2023' -> Dec 10
    20: datetime(2024, 7, 21),   # '21/07/2024' -> Jul 21
    21: datetime(2023, 9, 4),    # '04/09/2023' -> Sep 4
    22: datetime(2024, 2, 16),   # '16/02/2024' -> Feb 16
    23: datetime(2024, 5, 27),   # '27/05/2024' -> May 27
    24: datetime(2024, 8, 9),    # '09/08/2024' -> Aug 9
    25: datetime(2023, 11, 1),   # '01/11/2023' -> Nov 1
    26: datetime(2024, 4, 13),   # '13/04/2024' -> Apr 13
    27: datetime(2024, 6, 22),   # '22/06/2024' -> Jun 22
    28: datetime(2024, 1, 7),    # '07/01/2024' -> Jan 7
    29: datetime(2023, 10, 19),  # '19/10/2023' -> Oct 19
    30: datetime(2024, 3, 11),   # '11/03/2024' -> Mar 11
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Orders'] if 'Orders' in wb.sheetnames else wb.active

    # Component 1: All D2:D30 cells are datetime objects, not text strings (0.40 points)
    # Initial state has text strings; golden state has datetime objects.
    try:
        datetime_count = 0
        total_cells = 29  # D2:D30
        for row in range(2, 31):
            val = ws.cell(row=row, column=4).value
            if isinstance(val, datetime):
                datetime_count += 1

        if datetime_count == total_cells:
            print(f"PASS: Component 1 -- All {total_cells} cells in D2:D30 are datetime objects (0.40 pts)")
            total_score += 0.40
        elif datetime_count > 0:
            # Partial credit proportional to how many are converted
            partial = 0.40 * (datetime_count / total_cells)
            print(f"PARTIAL: Component 1 -- {datetime_count}/{total_cells} cells are datetime objects ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- No cells in D2:D30 are datetime objects (0/29 converted)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Dates correctly reflect DD/MM/YYYY European interpretation (0.35 points)
    # The key change: day and month must be swapped from MM/DD/YYYY to DD/MM/YYYY.
    # e.g., '05/02/2024' must become Feb 5 (not May 2)
    try:
        correct_count = 0
        for row in range(2, 31):
            val = ws.cell(row=row, column=4).value
            if isinstance(val, datetime):
                expected = EXPECTED_DATES[row]
                if val.year == expected.year and val.month == expected.month and val.day == expected.day:
                    correct_count += 1
                else:
                    print(f"  MISMATCH D{row}: got {val.strftime('%Y-%m-%d')}, expected {expected.strftime('%Y-%m-%d')}")

        if correct_count == 29:
            print(f"PASS: Component 2 -- All 29 dates correctly reflect DD/MM/YYYY interpretation (0.35 pts)")
            total_score += 0.35
        elif correct_count > 0:
            partial = 0.35 * (correct_count / 29)
            print(f"PARTIAL: Component 2 -- {correct_count}/29 dates have correct European interpretation ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- No dates correctly reflect DD/MM/YYYY interpretation")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Number format is a date format, not text '@' (0.25 points)
    # Initial state uses '@' (text format); golden state should use a date format.
    try:
        date_format_count = 0
        for row in range(2, 31):
            nf = ws.cell(row=row, column=4).number_format
            # A date format should contain date-related tokens (d, m, y) or be 'General'
            # but NOT be '@' (text format) or 'General' (which would mean no explicit date format)
            if nf and nf != '@' and nf != 'General':
                # Check it looks like a date format (contains d, m, or y patterns)
                nf_lower = nf.lower()
                if 'd' in nf_lower or 'm' in nf_lower or 'y' in nf_lower:
                    date_format_count += 1

        if date_format_count == 29:
            print(f"PASS: Component 3 -- All 29 cells have a date number format (0.25 pts)")
            total_score += 0.25
        elif date_format_count > 0:
            partial = 0.25 * (date_format_count / 29)
            print(f"PARTIAL: Component 3 -- {date_format_count}/29 cells have date number format ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- No cells have a date number format (still text '@' or General)")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
