"""
Initial Setup: Q3 Report sheet protected with password 'report123'
Task ID: calc_adv_protect_unprotect_013
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_protect_unprotect_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Q3 Report (password-protected) ---
    ws1 = wb.active
    ws1.title = 'Q3 Report'

    # Headers
    headers = ['Category', 'Q3 2025 ($)', 'Q2 2025 ($)', 'Change (%)', 'Notes']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    ws1.row_dimensions[1].height = 22

    # Financial data rows
    data = [
        ['Revenue',          2_847_320,  2_614_890,  8.9,   'Strong demand across all segments'],
        ['Cost of Goods',    1_203_440,  1_098_560,  9.5,   'Raw material costs elevated'],
        ['Gross Profit',     1_643_880,  1_516_330,  8.4,   '=IFERROR((C3-D3)/ABS(D3)*100,0)'],
        ['Operating Expenses',  512_760,   487_340,  5.2,   'Headcount increase in Q3'],
        ['EBITDA',           1_131_120,  1_028_990,  9.9,   'Margin improvement vs Q2'],
        ['Depreciation',        87_500,    87_500,   0.0,   'Straight-line, unchanged'],
        ['EBIT',             1_043_620,   941_490,  10.8,   'Strong operational leverage'],
        ['Interest Expense',    34_200,    36_450,  -6.2,   'Refinanced in July 2025'],
        ['Pre-Tax Income',   1_009_420,   905_040,  11.5,   ''],
        ['Income Tax',         302_826,   271_512,  11.5,   '30% effective rate'],
        ['Net Income',         706_594,   633_528,  11.5,   'Record quarter'],
        ['EPS (diluted)',           2.94,     2.64,  11.4,  'Based on 240.3M diluted shares'],
    ]

    value_font = Font(size=10, name='Calibri')
    center_align = Alignment(horizontal='center')
    right_align = Alignment(horizontal='right')

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 1:
                cell.font = Font(size=10, name='Calibri', bold=True)
                cell.alignment = Alignment(horizontal='left')
            elif c in (2, 3):
                cell.number_format = '#,##0'
                cell.alignment = right_align
                cell.font = value_font
            elif c == 4:
                cell.number_format = '0.0'
                cell.alignment = center_align
                cell.font = value_font
            else:
                cell.alignment = Alignment(horizontal='left')
                cell.font = value_font

        # Alternate row shading
        if r % 2 == 0:
            for c in range(1, 6):
                cell = ws1.cell(row=r, column=c)
                if c != 1 or not ws1.cell(row=r, column=c).fill.fgColor.rgb.startswith('FF44'):
                    cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    # Column widths
    ws1.column_dimensions['A'].width = 24
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 16
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 38

    # Summary labels below data
    ws1.cell(row=14, column=1, value='Report Period:').font = Font(bold=True, size=10)
    ws1.cell(row=14, column=2, value='July 1 – September 30, 2025').font = Font(size=10)
    ws1.cell(row=15, column=1, value='Prepared by:').font = Font(bold=True, size=10)
    ws1.cell(row=15, column=2, value='Finance Dept — Angela Torres').font = Font(size=10)
    ws1.cell(row=16, column=1, value='Status:').font = Font(bold=True, size=10)
    ws1.cell(row=16, column=2, value='FINAL — Approved by CFO').font = Font(size=10)

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # APPLY SHEET PROTECTION with password 'report123'
    ws1.protection.sheet = True
    ws1.protection.password = 'report123'
    ws1.protection.enable()

    # --- Sheet 2: Annual Summary ---
    ws2 = wb.create_sheet('Annual Summary')
    ws2['A1'] = 'Fiscal Year 2025 — Running Totals'
    ws2['A1'].font = Font(bold=True, size=12, name='Calibri')

    ws2_headers = ['Quarter', 'Revenue ($)', 'Net Income ($)', 'EPS (diluted)']
    for col, h in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=10, name='Calibri')
        cell.fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    fy_data = [
        ['Q1 2025', 2_412_100, 581_250, 2.42],
        ['Q2 2025', 2_614_890, 633_528, 2.64],
        ['Q3 2025', 2_847_320, 706_594, 2.94],
        ['Q4 2025 (est.)', 2_950_000, 740_000, 3.08],
    ]
    for r, row_data in enumerate(fy_data, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 1:
                cell.font = Font(size=10, name='Calibri')
            elif c in (2, 3):
                cell.number_format = '#,##0'
                cell.font = Font(size=10, name='Calibri')
                cell.alignment = Alignment(horizontal='right')
            elif c == 4:
                cell.number_format = '0.00'
                cell.font = Font(size=10, name='Calibri')
                cell.alignment = Alignment(horizontal='center')

    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 14

    # --- Sheet 3: Assumptions ---
    ws3 = wb.create_sheet('Assumptions')
    ws3['A1'] = 'Modeling Assumptions — Q3 2025'
    ws3['A1'].font = Font(bold=True, size=12, name='Calibri')

    assumptions = [
        ['Tax Rate', '30.0%', 'Based on current applicable jurisdiction rates'],
        ['Diluted Shares', '240.3M', 'As reported in Q3 10-Q filing'],
        ['Depreciation Method', 'Straight-line', '10-year asset life assumed'],
        ['FX Rate (USD/EUR)', '1.085', 'Average rate for Q3 2025'],
        ['FX Rate (USD/GBP)', '1.272', 'Average rate for Q3 2025'],
        ['Inflation Assumption', '3.2%', 'CPI-based estimate'],
    ]
    for r, row_data in enumerate(assumptions, 3):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = Font(size=10, name='Calibri')
            if r == 3:
                cell.font = Font(bold=True, size=10, name='Calibri')

    ws3.column_dimensions['A'].width = 24
    ws3.column_dimensions['B'].width = 16
    ws3.column_dimensions['C'].width = 44

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet protection status:')
    wb2 = openpyxl.load_workbook(OUTPUT)
    for sheet in wb2.sheetnames:
        ws = wb2[sheet]
        print(f'  {sheet}: protected={ws.protection.sheet}')


create_initial()
