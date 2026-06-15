"""
Reward Script: Unprotect the 'Q3 Report' sheet
Task ID: calc_adv_protect_unprotect_013
Domain: libreoffice_calc
Scoring:
  Component 1: 'Q3 Report' sheet protection is disabled (0.7 pts)
  Component 2: No password is set on the 'Q3 Report' sheet (0.3 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_protect_unprotect_013'
TARGET_SHEET = 'Q3 Report'


def verify_task(file_path):
    """
    Verify that the 'Q3 Report' sheet has been unprotected.
    The initial file has protection.sheet=True with password '9582'.
    The golden file must have protection.sheet=False and no password.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Target sheet must exist
    if TARGET_SHEET not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{TARGET_SHEET}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[TARGET_SHEET]

    # Component 1: Sheet protection is disabled (0.7 points)
    # Initial state: protection.sheet == True
    # Golden state:  protection.sheet == False
    try:
        protection_enabled = ws.protection.sheet
        if not protection_enabled:
            print(f"PASS: Component 1 — Sheet '{TARGET_SHEET}' protection is disabled (protection.sheet=False) (0.7 pts)")
            total_score += 0.7
        else:
            print(f"FAIL: Component 1 — Sheet '{TARGET_SHEET}' is still protected (protection.sheet=True)")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check sheet protection: {e}")

    # Component 2: No password is set on the sheet (0.3 points)
    # Initial state: protection.password == '9582' (hashed value)
    # Golden state:  protection.password == None
    try:
        password_value = ws.protection.password
        if password_value is None:
            print(f"PASS: Component 2 — Sheet '{TARGET_SHEET}' has no password set (protection.password=None) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Sheet '{TARGET_SHEET}' still has a password set (protection.password={repr(password_value)})")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check sheet password: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
