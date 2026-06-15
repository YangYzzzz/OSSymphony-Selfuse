"""
Reward Script: Extract PDF authors and create sorted spreadsheet
Task ID: osworld_multi_apps_pdf_author_extract_004
Domain: libreoffice_calc
Scoring:
  - Component 1: Correct headers in row 1 (Name, Email, Institution)           0.2 pts
  - Component 2: Correct number of data rows (6 rows for 6 PDFs)               0.2 pts
  - Component 3: All author data correct (Name, Email, Institution match)       0.4 pts
  - Component 4: Rows sorted alphabetically by author last name                 0.2 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_author_extract_004'

# Ground truth author data extracted from the 6 PDF papers
# These are the exact values from the golden file
EXPECTED_AUTHORS = [
    {'name': 'David Chen',          'email': 'd.chen@mit.edu',            'institution': 'MIT CSAIL'},
    {'name': 'Emma Fischer',        'email': 'e.fischer@stanford.edu',    'institution': 'Stanford University'},
    {'name': 'James Huang',         'email': 'j.huang@berkeley.edu',      'institution': 'UC Berkeley'},
    {'name': 'Priya Kumar',         'email': 'p.kumar@cmu.edu',           'institution': 'Carnegie Mellon University'},
    {'name': 'Alexander Nakamura',  'email': 'a.nakamura@caltech.edu',    'institution': 'California Institute of Technology'},
    {'name': 'Sofia Rodriguez',     'email': 's.rodriguez@utexas.edu',    'institution': 'University of Texas at Austin'},
]

# The expected last names in sorted order
EXPECTED_LAST_NAMES_SORTED = ['Chen', 'Fischer', 'Huang', 'Kumar', 'Nakamura', 'Rodriguez']


def author_matches(actual, expected):
    """Return True if actual author dict matches expected author dict (case-insensitive)."""
    name_match = actual['name'].lower() == expected['name'].lower()
    email_match = actual['email'].lower() == expected['email'].lower()
    inst_match = actual['institution'].lower() == expected['institution'].lower()
    return name_match and email_match and inst_match


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Correct headers in row 1 (0.2 points)
    # Headers must be exactly 'Name', 'Email', 'Institution' in columns A, B, C
    try:
        header_a = ws.cell(row=1, column=1).value
        header_b = ws.cell(row=1, column=2).value
        header_c = ws.cell(row=1, column=3).value

        expected_headers = ('Name', 'Email', 'Institution')
        actual_headers = (
            str(header_a).strip() if header_a else '',
            str(header_b).strip() if header_b else '',
            str(header_c).strip() if header_c else '',
        )

        # Case-insensitive comparison for robustness
        headers_correct = (
            actual_headers[0].lower() == 'name' and
            actual_headers[1].lower() == 'email' and
            actual_headers[2].lower() == 'institution'
        )
        if headers_correct:
            print(f"PASS: Component 1 — Headers correct: {actual_headers} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — Expected headers {expected_headers}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct number of data rows (0.2 points)
    # Must have exactly 6 data rows (one per PDF paper), plus the header row
    try:
        # Count non-empty data rows (excluding header)
        data_row_count = 0
        for row_idx in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=1).value
            if name_val is not None and str(name_val).strip() != '':
                data_row_count += 1

        if data_row_count == 6:
            print(f"PASS: Component 2 — Correct data row count: {data_row_count} rows (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Expected 6 data rows, found {data_row_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All author data correct (Name, Email, Institution) (0.4 points)
    # Check that all expected authors appear in the spreadsheet with partial credit per author
    try:
        # Collect all data rows from the file
        file_authors = []
        for row_idx in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=1).value
            email_val = ws.cell(row=row_idx, column=2).value
            inst_val = ws.cell(row=row_idx, column=3).value
            if name_val is not None and str(name_val).strip() != '':
                file_authors.append({
                    'name': str(name_val).strip(),
                    'email': str(email_val).strip() if email_val else '',
                    'institution': str(inst_val).strip() if inst_val else '',
                })

        # Check how many expected authors are correctly represented
        matched_count = 0
        for expected in EXPECTED_AUTHORS:
            # Check if any actual row matches this expected author
            matched = any(author_matches(actual, expected) for actual in file_authors)
            if matched:
                matched_count += 1
            else:
                print(f"FAIL: Component 3 — Author not found or mismatched: {expected['name']}")

        # Award partial credit: 0.4 * (matched / 6)
        component3_score = round((matched_count / 6) * 0.4, 4)
        if matched_count == 6:
            print(f"PASS: Component 3 — All 6 authors correctly recorded ({component3_score} pts)")
            total_score += component3_score
        elif matched_count > 0:
            print(f"PARTIAL: Component 3 — {matched_count}/6 authors matched ({component3_score} pts)")
            total_score += component3_score
        else:
            print(f"FAIL: Component 3 — No authors matched (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Rows sorted alphabetically by author last name (0.2 points)
    # The rows must be sorted A-Z by the author's last name
    try:
        # Collect data row names in file order
        file_names_in_order = []
        for row_idx in range(2, ws.max_row + 1):
            name_val = ws.cell(row=row_idx, column=1).value
            if name_val is not None and str(name_val).strip() != '':
                file_names_in_order.append(str(name_val).strip())

        # Extract last names
        def get_last_name(full_name):
            parts = full_name.strip().split()
            return parts[-1] if parts else ''

        actual_last_names = [get_last_name(n) for n in file_names_in_order]

        # Check if the last names are in ascending alphabetical order
        is_sorted = (
            len(actual_last_names) > 1 and
            all(
                actual_last_names[i].lower() <= actual_last_names[i + 1].lower()
                for i in range(len(actual_last_names) - 1)
            )
        )

        if is_sorted:
            print(f"PASS: Component 4 — Rows sorted by last name: {actual_last_names} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Rows not sorted by last name. Found order: {actual_last_names}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 1)}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/survey_authors.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
