"""
Reward Script: Energy consumption analysis with formulas, YoY comparisons, charts, and conditional formatting
Task ID: calc_wf_069
Domain: libreoffice_calc
Scoring:
  Component 1: Cost formulas in D/G/J columns (0.20)
  Component 2: Total Cost formulas in K column (0.15)
  Component 3: YoY Change % column L with formulas (0.20)
  Component 4: Average monthly formula in K27 (0.10)
  Component 5: Charts - line chart + stacked bar chart (0.20)
  Component 6: Conditional formatting on K14:K25 (0.15)
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_069'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Utilities' sheet must exist
    if 'Utilities' not in wb.sheetnames:
        print("FAIL: 'Utilities' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Utilities']

    # Component 1: Cost formulas in D, G, J columns (0.20 points)
    # Task requires: Elec Cost = kWh * Rate (D=B*C), Gas Cost = therms * Rate (G=E*F), Water Cost = gal * Rate (J=H*I)
    # Initial file has hardcoded values; golden has formulas.
    try:
        formula_count = 0
        expected_formulas = 0
        for r in range(2, 26):  # rows 2-25 (24 months)
            expected_formulas += 3
            d_val = ws.cell(r, 4).value  # Elec Cost
            g_val = ws.cell(r, 7).value  # Gas Cost
            j_val = ws.cell(r, 10).value  # Water Cost

            # Check if they are formulas referencing the correct cells
            if isinstance(d_val, str) and '=' in d_val:
                # Should be something like =B2*C2
                if re.search(r'=\s*B\d+\s*\*\s*C\d+', d_val, re.IGNORECASE):
                    formula_count += 1
            if isinstance(g_val, str) and '=' in g_val:
                if re.search(r'=\s*E\d+\s*\*\s*F\d+', g_val, re.IGNORECASE):
                    formula_count += 1
            if isinstance(j_val, str) and '=' in j_val:
                if re.search(r'=\s*H\d+\s*\*\s*I\d+', j_val, re.IGNORECASE):
                    formula_count += 1

        if formula_count >= expected_formulas * 0.8:  # allow some tolerance
            print(f"PASS: Component 1 - Cost formulas in D/G/J ({formula_count}/{expected_formulas} correct) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 - Cost formulas in D/G/J: only {formula_count}/{expected_formulas} are formulas")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Total Cost formulas in K column (0.15 points)
    # Task requires: monthly total = sum of all utility costs (K = D + G + J)
    # Initial has hardcoded totals; golden has formulas.
    try:
        total_formula_count = 0
        expected_total = 24  # rows 2-25
        for r in range(2, 26):
            k_val = ws.cell(r, 11).value  # Total Cost
            if isinstance(k_val, str) and '=' in k_val:
                # Should reference D, G, J in some form (=D2+G2+J2 or =SUM(D2,G2,J2) etc.)
                k_upper = k_val.upper().replace(' ', '')
                if ('D' in k_upper and 'G' in k_upper and 'J' in k_upper):
                    total_formula_count += 1

        if total_formula_count >= expected_total * 0.8:
            print(f"PASS: Component 2 - Total Cost formulas in K ({total_formula_count}/{expected_total}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 - Total Cost formulas in K: only {total_formula_count}/{expected_total} are formulas")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: YoY Change % column L with header and formulas (0.20 points)
    # Task requires: year-over-year comparisons for 2025 months referencing 2024 months
    # Initial file has no column L header or data. Golden has L1="YoY Change %" and L14:L25 with formulas.
    try:
        yoy_score = 0.0
        # Check L1 header
        l1_val = ws.cell(1, 12).value
        if l1_val and 'yoy' in str(l1_val).lower().replace('-', '').replace(' ', ''):
            yoy_score += 0.05
            print(f"  PASS: YoY header found: '{l1_val}'")
        elif l1_val and 'year' in str(l1_val).lower():
            yoy_score += 0.05
            print(f"  PASS: YoY header found: '{l1_val}'")
        else:
            print(f"  FAIL: YoY header in L1: found '{l1_val}'")

        # Check L14:L25 for YoY formulas (should reference current year K vs previous year K)
        yoy_formula_count = 0
        for r in range(14, 26):
            l_val = ws.cell(r, 12).value
            if isinstance(l_val, str) and '=' in l_val:
                l_upper = l_val.upper().replace(' ', '')
                # Should reference K{r} and K{r-12} in some form
                prev_row = r - 12
                if f'K{r}' in l_upper and f'K{prev_row}' in l_upper:
                    yoy_formula_count += 1

        if yoy_formula_count >= 10:  # at least 10 of 12 months
            yoy_score += 0.15
            print(f"  PASS: YoY formulas found ({yoy_formula_count}/12)")
        else:
            print(f"  FAIL: YoY formulas: only {yoy_formula_count}/12 found")

        if yoy_score > 0:
            print(f"PASS: Component 3 - YoY Change % ({yoy_score} pts)")
            total_score += yoy_score
        else:
            print(f"FAIL: Component 3 - No YoY data found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Average monthly formula (0.10 points)
    # Task requires: average monthly calculation for current year costs
    # Initial has no average formula. Golden has AVERAGE formula in K27.
    try:
        avg_found = False
        # Search rows 26-35 for an AVERAGE formula referencing K column
        for r in range(26, 36):
            for c in range(1, 17):
                val = ws.cell(r, c).value
                if isinstance(val, str) and '=' in val:
                    val_upper = val.upper().replace(' ', '')
                    if 'AVERAGE' in val_upper and 'K' in val_upper:
                        avg_found = True
                        print(f"  Found AVERAGE formula at row {r}, col {c}: {val}")
                        break
            if avg_found:
                break

        if avg_found:
            print(f"PASS: Component 4 - Average monthly formula found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 - No AVERAGE formula for monthly costs found")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Charts (0.20 points)
    # Task requires: line chart of monthly totals (both years overlaid) and stacked bar of cost breakdown
    # Initial has 0 charts. Golden has 2 charts.
    try:
        charts = ws._charts
        num_charts = len(charts)

        if num_charts >= 2:
            chart_score = 0.0
            has_line = False
            has_bar_stacked = False

            for ch in charts:
                ch_class = ch.__class__.__name__
                if ch_class == 'LineChart':
                    has_line = True
                    # Should have 2 series (2024 and 2025 totals)
                    if len(ch.series) >= 2:
                        chart_score += 0.10
                        print(f"  PASS: Line chart with {len(ch.series)} series (both years)")
                    else:
                        chart_score += 0.05
                        print(f"  PARTIAL: Line chart found but only {len(ch.series)} series (expected >= 2)")
                elif ch_class == 'BarChart':
                    # Check if stacked
                    if hasattr(ch, 'grouping') and ch.grouping == 'stacked':
                        has_bar_stacked = True
                        if len(ch.series) >= 3:
                            chart_score += 0.10
                            print(f"  PASS: Stacked bar chart with {len(ch.series)} series (cost breakdown)")
                        else:
                            chart_score += 0.05
                            print(f"  PARTIAL: Stacked bar chart but only {len(ch.series)} series (expected >= 3)")
                    else:
                        # Bar chart but not stacked
                        chart_score += 0.05
                        print(f"  PARTIAL: Bar chart found but not stacked (grouping={getattr(ch, 'grouping', 'N/A')})")

            if not has_line:
                print(f"  FAIL: No line chart found")
            if not has_bar_stacked:
                print(f"  FAIL: No stacked bar chart found")

            total_score += chart_score
            print(f"PASS: Component 5 - Charts ({chart_score} pts)")
        elif num_charts == 1:
            total_score += 0.05
            print(f"PARTIAL: Component 5 - Only 1 chart found (expected 2) (0.05 pts)")
        else:
            print(f"FAIL: Component 5 - No charts found (expected 2)")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Conditional formatting on months exceeding average (0.15 points)
    # Task requires: highlight months where total > average (orange fill)
    # Initial has 0 CF rules. Golden has CF on K14:K25 with orange fill.
    try:
        cf_rules = list(ws.conditional_formatting)
        if len(cf_rules) > 0:
            cf_score = 0.0
            found_relevant_cf = False
            for cf in cf_rules:
                cf_range = str(cf)
                for rule in cf.rules:
                    # Check if it references AVERAGE or K column for comparison
                    rule_formulas = rule.formula if rule.formula else []
                    for f in rule_formulas:
                        f_upper = str(f).upper().replace(' ', '')
                        if 'AVERAGE' in f_upper and 'K' in f_upper:
                            found_relevant_cf = True
                            cf_score = 0.15
                            # Check for orange-ish fill
                            if rule.dxf and rule.dxf.fill:
                                try:
                                    fill_color = rule.dxf.fill.fgColor.rgb
                                    print(f"  CF range: {cf_range}, formula: {f}, fill color: {fill_color}")
                                except:
                                    print(f"  CF range: {cf_range}, formula: {f}")
                            else:
                                print(f"  CF range: {cf_range}, formula: {f}")
                            break
                    if found_relevant_cf:
                        break
                if found_relevant_cf:
                    break

            if not found_relevant_cf:
                # Looser check: any CF on K column range
                for cf in cf_rules:
                    cf_range = str(cf)
                    if 'K' in cf_range.upper():
                        found_relevant_cf = True
                        cf_score = 0.10
                        print(f"  PARTIAL: CF on K column range: {cf_range}")
                        break

            if found_relevant_cf:
                total_score += cf_score
                print(f"PASS: Component 6 - Conditional formatting ({cf_score} pts)")
            else:
                print(f"FAIL: Component 6 - No relevant conditional formatting found")
        else:
            print(f"FAIL: Component 6 - No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
