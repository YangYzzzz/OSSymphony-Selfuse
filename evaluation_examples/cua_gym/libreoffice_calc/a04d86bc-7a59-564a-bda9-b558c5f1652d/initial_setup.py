"""
Initial Setup: Energy consumption analysis with 24 months of utility bills
Task ID: calc_wf_069
Domain: libreoffice_calc

Creates a spreadsheet with raw utility data (electricity, gas, water) for
24 months (12 current year + 12 previous year). No formulas, no charts,
no conditional formatting -- those are the task for the agent.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Utilities ---
    ws = wb.active
    ws.title = 'Utilities'

    # Headers
    headers = [
        'Month', 'Electricity kWh', 'Elec Rate', 'Elec Cost',
        'Gas Therms', 'Gas Rate', 'Gas Cost',
        'Water Gal', 'Water Rate', 'Water Cost', 'Total Cost'
    ]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Previous year data (2024) - rows 2-13
    prev_year_data = [
        ['Jan 2024', 920, 0.12, 110.40, 85, 1.05, 89.25, 4200, 0.005, 21.00, 220.65],
        ['Feb 2024', 880, 0.12, 105.60, 92, 1.05, 96.60, 3800, 0.005, 19.00, 221.20],
        ['Mar 2024', 810, 0.12, 97.20, 68, 1.05, 71.40, 4500, 0.005, 22.50, 191.10],
        ['Apr 2024', 750, 0.13, 97.50, 45, 1.08, 48.60, 5200, 0.006, 31.20, 177.30],
        ['May 2024', 830, 0.13, 107.90, 30, 1.08, 32.40, 6100, 0.006, 36.60, 176.90],
        ['Jun 2024', 1050, 0.13, 136.50, 18, 1.08, 19.44, 7200, 0.006, 43.20, 199.14],
        ['Jul 2024', 1280, 0.14, 179.20, 12, 1.10, 13.20, 8100, 0.007, 56.70, 249.10],
        ['Aug 2024', 1350, 0.14, 189.00, 10, 1.10, 11.00, 8500, 0.007, 59.50, 259.50],
        ['Sep 2024', 1100, 0.14, 154.00, 22, 1.10, 24.20, 6800, 0.007, 47.60, 225.80],
        ['Oct 2024', 870, 0.13, 113.10, 48, 1.08, 51.84, 5100, 0.006, 30.60, 195.54],
        ['Nov 2024', 900, 0.12, 108.00, 72, 1.05, 75.60, 4000, 0.005, 20.00, 203.60],
        ['Dec 2024', 960, 0.12, 115.20, 90, 1.05, 94.50, 3600, 0.005, 18.00, 227.70],
    ]

    # Current year data (2025) - rows 14-25
    curr_year_data = [
        ['Jan 2025', 950, 0.13, 123.50, 88, 1.10, 96.80, 4100, 0.006, 24.60, 244.90],
        ['Feb 2025', 910, 0.13, 118.30, 95, 1.10, 104.50, 3700, 0.006, 22.20, 245.00],
        ['Mar 2025', 840, 0.13, 109.20, 70, 1.10, 77.00, 4600, 0.006, 27.60, 213.80],
        ['Apr 2025', 780, 0.14, 109.20, 48, 1.12, 53.76, 5400, 0.007, 37.80, 200.76],
        ['May 2025', 860, 0.14, 120.40, 32, 1.12, 35.84, 6300, 0.007, 44.10, 200.34],
        ['Jun 2025', 1100, 0.14, 154.00, 20, 1.12, 22.40, 7500, 0.007, 52.50, 228.90],
        ['Jul 2025', 1320, 0.15, 198.00, 14, 1.15, 16.10, 8400, 0.008, 67.20, 281.30],
        ['Aug 2025', 1400, 0.15, 210.00, 11, 1.15, 12.65, 8800, 0.008, 70.40, 293.05],
        ['Sep 2025', 1140, 0.15, 171.00, 25, 1.15, 28.75, 7000, 0.008, 56.00, 255.75],
        ['Oct 2025', 900, 0.14, 126.00, 52, 1.12, 58.24, 5300, 0.007, 37.10, 221.34],
        ['Nov 2025', 930, 0.13, 120.90, 76, 1.10, 83.60, 4100, 0.006, 24.60, 229.10],
        ['Dec 2025', 990, 0.13, 128.70, 94, 1.10, 103.40, 3500, 0.006, 21.00, 253.10],
    ]

    all_data = prev_year_data + curr_year_data
    for r, row_data in enumerate(all_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format rate columns as decimal
    for row in range(2, 26):
        ws.cell(row=row, column=3).number_format = '$0.00'   # Elec Rate
        ws.cell(row=row, column=6).number_format = '$0.00'   # Gas Rate
        ws.cell(row=row, column=9).number_format = '$0.000'  # Water Rate
        # Format cost columns as currency
        ws.cell(row=row, column=4).number_format = '$#,##0.00'  # Elec Cost
        ws.cell(row=row, column=7).number_format = '$#,##0.00'  # Gas Cost
        ws.cell(row=row, column=10).number_format = '$#,##0.00' # Water Cost
        ws.cell(row=row, column=11).number_format = '$#,##0.00' # Total Cost

    # Adjust column widths
    col_widths = {'A': 14, 'B': 16, 'C': 12, 'D': 12, 'E': 14, 'F': 12,
                  'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
