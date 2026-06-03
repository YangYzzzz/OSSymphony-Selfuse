"""
Initial Setup: Elegant AutoFormat with custom header task
Task ID: calc_gg3_037
Domain: libreoffice_calc

Creates /home/user/formatted_report.xlsx with a 'Data' sheet containing
a 15-row x 5-column table (headers + 14 data rows) with NO formatting.
Opens the file in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
OUTPUT = f'{WORKDIR}/formatted_report.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    """Create a plain (unformatted) spreadsheet with product data."""
    try:
        import openpyxl
    except ImportError:
        subprocess.check_call(['pip3', 'install', 'openpyxl'])
        import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Data'

    # Headers in row 1: A1:E1
    headers = ['Product', 'Q1', 'Q2', 'Q3', 'Total']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 14 data rows (rows 2-15) with realistic product data
    products = [
        ['Wireless Mouse',        1250, 1480, 1320, None],
        ['Mechanical Keyboard',    890,  920, 1050, None],
        ['USB-C Hub',             2100, 2350, 2180, None],
        ['Monitor Stand',          670,  710,  690, None],
        ['Webcam HD Pro',         1540, 1620, 1780, None],
        ['Noise-Cancel Headset',   430,  510,  480, None],
        ['Laptop Sleeve 15"',      980, 1040, 1100, None],
        ['Desk Lamp LED',          320,  360,  340, None],
        ['Ergonomic Chair',        150,  180,  210, None],
        ['Portable SSD 1TB',      1870, 2010, 1930, None],
        ['Bluetooth Speaker',      760,  820,  790, None],
        ['Charging Pad',          1100, 1250, 1180, None],
        ['Cable Management Kit',   540,  580,  620, None],
        ['Screen Protector',      2400, 2550, 2680, None],
    ]

    for r, row_data in enumerate(products, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Product name
        ws.cell(row=r, column=2, value=row_data[1])  # Q1
        ws.cell(row=r, column=3, value=row_data[2])  # Q2
        ws.cell(row=r, column=4, value=row_data[3])  # Q3
        # Total = Q1 + Q2 + Q3
        total = row_data[1] + row_data[2] + row_data[3]
        ws.cell(row=r, column=5, value=total)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


# 1. Create the spreadsheet (no formatting)
create_initial()

# 2. Launch LibreOffice Calc with the file
launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')
