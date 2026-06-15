"""
Reward script for calc_gg3_037:
Verify that the 'Elegant' AutoFormat was applied to A1:E15 and the header row
(A1:E1) has dark green background (#006400) with white bold text.

Scoring rubric (total = 1.0):
  1. Header background is dark green FF006400           — 0.30
  2. Header font color is white FFFFFFFF                — 0.20
  3. Header font is bold                                — 0.15
  4. Data rows have Elegant AutoFormat styling (fills)  — 0.20
  5. Data rows have borders (Elegant AutoFormat)        — 0.15
"""

import openpyxl

FILE_PATH = "/home/user/formatted_report.xlsx"

EXPECTED_HEADERS = ["Product", "Q1", "Q2", "Q3", "Total"]

score = 0.0

try:
    wb = openpyxl.load_workbook(FILE_PATH)
except Exception as e:
    print(f"ERROR: Cannot open workbook: {e}")
    print(f"REWARD: 0.0")
    raise SystemExit(0)

try:
    ws = wb["Data"]
except Exception as e:
    print(f"ERROR: 'Data' sheet not found: {e}")
    print(f"REWARD: 0.0")
    raise SystemExit(0)

# ── Component 1: Header background is dark green (0.30) ──
try:
    green_count = 0
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        fill_fg = None
        try:
            if cell.fill.fill_type == "solid" and cell.fill.fgColor:
                fill_fg = cell.fill.fgColor.rgb
        except Exception:
            pass
        if fill_fg and str(fill_fg).upper() == "FF006400":
            green_count += 1
    component1 = (green_count / 5) * 0.30
    score += component1
    print(f"Component 1 (header bg green): {green_count}/5 cells correct -> {component1:.4f}")
except Exception as e:
    print(f"Component 1 ERROR: {e}")

# ── Component 2: Header font color is white (0.20) ──
try:
    white_count = 0
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        font_color = None
        try:
            font_color = cell.font.color.rgb
        except Exception:
            pass
        if font_color and str(font_color).upper() == "FFFFFFFF":
            white_count += 1
    component2 = (white_count / 5) * 0.20
    score += component2
    print(f"Component 2 (header font white): {white_count}/5 cells correct -> {component2:.4f}")
except Exception as e:
    print(f"Component 2 ERROR: {e}")

# ── Component 3: Header font is bold (0.15) ──
try:
    bold_count = 0
    for c in range(1, 6):
        cell = ws.cell(row=1, column=c)
        if cell.font.bold:
            bold_count += 1
    component3 = (bold_count / 5) * 0.15
    score += component3
    print(f"Component 3 (header bold): {bold_count}/5 cells correct -> {component3:.4f}")
except Exception as e:
    print(f"Component 3 ERROR: {e}")

# ── Component 4: Data rows have Elegant AutoFormat styling (0.20) ──
# Elegant AutoFormat applies alternating row fills: white (FFFFFFFF) for even rows,
# light gray (FFD4D4D4) for odd rows (row 2=even-index=white, row 3=odd-index=gray, etc.)
try:
    styled_cells = 0
    total_data_cells = 14 * 5  # rows 2-15, cols 1-5 = 70 cells
    for r in range(2, 16):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            try:
                if cell.fill.fill_type == "solid" and cell.fill.fgColor:
                    fg = str(cell.fill.fgColor.rgb).upper()
                    # Even rows (2,4,6,...) should be white, odd rows (3,5,7,...) should be gray
                    if r % 2 == 0 and fg == "FFFFFFFF":
                        styled_cells += 1
                    elif r % 2 == 1 and fg == "FFD4D4D4":
                        styled_cells += 1
            except Exception:
                pass
    component4 = (styled_cells / total_data_cells) * 0.20
    score += component4
    print(f"Component 4 (data row styling): {styled_cells}/{total_data_cells} cells correct -> {component4:.4f}")
except Exception as e:
    print(f"Component 4 ERROR: {e}")

# ── Component 5: Data rows have borders from Elegant AutoFormat (0.15) ──
try:
    bordered_cells = 0
    total_border_checks = 15 * 5  # rows 1-15, cols 1-5 = 75 cells
    for r in range(1, 16):
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c)
            try:
                # Check that at least some border style is applied
                has_border = False
                if cell.border.left.style is not None:
                    has_border = True
                elif cell.border.right.style is not None:
                    has_border = True
                elif cell.border.top.style is not None:
                    has_border = True
                elif cell.border.bottom.style is not None:
                    has_border = True
                if has_border:
                    bordered_cells += 1
            except Exception:
                pass
    component5 = (bordered_cells / total_border_checks) * 0.15
    score += component5
    print(f"Component 5 (borders): {bordered_cells}/{total_border_checks} cells with borders -> {component5:.4f}")
except Exception as e:
    print(f"Component 5 ERROR: {e}")

# Round to avoid floating point issues
score = round(score, 2)
# Clamp to [0.0, 1.0]
score = max(0.0, min(1.0, score))

print(f"REWARD: {score}")
