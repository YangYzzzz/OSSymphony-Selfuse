"""
Reward Script: Format cells E2:E30 with thousand separators and no decimal places
Task ID: calc_gfl_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): E2:E30 have thousand-separator number format
  Component 2 (0.3): Format has no decimal places
  Component 3 (0.2): Underlying numeric values unchanged
  Component 4 (0.1): Other columns retain General formatting
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_033'

# Expected raw values for E2:E30 (from the initial file)
EXPECTED_VALUES = [
    1250000, 875000, 3400000, 2150000, 1890000, 960000, 4500000, 2780000,
    5200000, 6100000, 8750000, 4300000, 3650000, 1980000, 7200000, 5400000,
    3100000, 2850000, 4100000, 3750000, 9200000, 6800000, 5600000, 3900000,
    2400000, 1750000, 4850000, 1620000, 7300000
]


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        ws = wb['Annual Revenue']
    except KeyError:
        print("CRITICAL: Sheet 'Annual Revenue' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: E2:E30 have thousand-separator number format (0.4 points)
    # The format must contain comma grouping (e.g., #,##0 or similar)
    try:
        formatted_count = 0
        total_cells = 29  # E2:E30
        for r in range(2, 31):
            cell = ws.cell(row=r, column=5)
            fmt = cell.number_format
            # Check that format contains comma grouping pattern
            if fmt and ',' in fmt and '#' in fmt:
                formatted_count += 1
        if formatted_count == total_cells:
            print(f"PASS: Component 1 — All {total_cells} cells E2:E30 have thousand-separator format (0.4 pts)")
            total_score += 0.4
        elif formatted_count > 0:
            partial = round(0.4 * formatted_count / total_cells, 2)
            print(f"PARTIAL: Component 1 — {formatted_count}/{total_cells} cells formatted ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in E2:E30 have thousand-separator format")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Format has no decimal places (0.3 points)
    # The format must NOT show decimal digits (no .0, .00, etc.)
    try:
        no_decimal_count = 0
        for r in range(2, 31):
            cell = ws.cell(row=r, column=5)
            fmt = cell.number_format
            # A format with decimals would contain a period followed by 0s
            # e.g., #,##0.00 has decimals, #,##0 does not
            if fmt and ',' in fmt and '.' not in fmt:
                no_decimal_count += 1
        if no_decimal_count == total_cells:
            print(f"PASS: Component 2 — All {total_cells} cells have no decimal places in format (0.3 pts)")
            total_score += 0.3
        elif no_decimal_count > 0:
            partial = round(0.3 * no_decimal_count / total_cells, 2)
            print(f"PARTIAL: Component 2 — {no_decimal_count}/{total_cells} cells without decimals ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Cells in E2:E30 either have decimals or wrong format")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Underlying numeric values are unchanged (0.2 points)
    # This is a compound check: values must match AND format must be applied
    # (If format is not applied, this component does not score — prevents initial scoring)
    try:
        # Gate: at least some formatting must be present (task-change anchor)
        has_any_formatting = any(
            ws.cell(row=r, column=5).number_format != 'General'
            for r in range(2, 31)
        )
        if not has_any_formatting:
            print("FAIL: Component 3 — No formatting applied, skipping value integrity check")
        else:
            values_match = 0
            for i, r in enumerate(range(2, 31)):
                cell = ws.cell(row=r, column=5)
                val = cell.value
                expected = EXPECTED_VALUES[i]
                if val is not None and isinstance(val, (int, float)):
                    if abs(val - expected) < 0.01:
                        values_match += 1
            if values_match == total_cells:
                print(f"PASS: Component 3 — All {total_cells} values unchanged after formatting (0.2 pts)")
                total_score += 0.2
            elif values_match > 0:
                partial = round(0.2 * values_match / total_cells, 2)
                print(f"PARTIAL: Component 3 — {values_match}/{total_cells} values intact ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 — Values in E2:E30 appear corrupted")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Other columns retain General formatting (0.1 points)
    # Compound check: format change applied to E AND other cols untouched
    try:
        has_any_formatting = any(
            ws.cell(row=r, column=5).number_format != 'General'
            for r in range(2, 31)
        )
        if not has_any_formatting:
            print("FAIL: Component 4 — No formatting applied to E, skipping other-column check")
        else:
            bad_format_cells = []
            for col_idx in [1, 2, 3, 4, 6, 7]:  # A, B, C, D, F, G
                for r in range(2, 31):
                    fmt = ws.cell(row=r, column=col_idx).number_format
                    if fmt != 'General':
                        bad_format_cells.append((col_idx, r, fmt))
            if len(bad_format_cells) == 0:
                print("PASS: Component 4 — Other columns retain General formatting (0.1 pts)")
                total_score += 0.1
            else:
                for col_idx, r, fmt in bad_format_cells[:3]:
                    print(f"  WARNING: Col {col_idx} row {r} has format '{fmt}' instead of General")
                print("FAIL: Component 4 — Some non-E columns have unexpected formatting")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'

# Persist unsaved state if LibreOffice is open
persist_app_state("libreoffice_calc")

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
