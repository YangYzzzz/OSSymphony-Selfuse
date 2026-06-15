"""
Initial Setup: Format cells E2:E30 with thousand separators
Task ID: calc_gfl_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Annual Revenue'

    # Headers
    headers = ['Division', 'Product Line', 'Region', 'Quarter', 'Revenue', 'Forecast', 'Variance']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 29 rows of realistic revenue data
    data = [
        ['Technology', 'Cloud Services', 'North America', 'Q1 2025', 1250000, 1300000, -50000],
        ['Technology', 'Cloud Services', 'Europe', 'Q1 2025', 875000, 900000, -25000],
        ['Technology', 'Enterprise Software', 'North America', 'Q1 2025', 3400000, 3200000, 200000],
        ['Technology', 'Enterprise Software', 'Asia Pacific', 'Q1 2025', 2150000, 2000000, 150000],
        ['Technology', 'Cybersecurity', 'North America', 'Q2 2025', 1890000, 1750000, 140000],
        ['Technology', 'Cybersecurity', 'Europe', 'Q2 2025', 960000, 1000000, -40000],
        ['Healthcare', 'Medical Devices', 'North America', 'Q1 2025', 4500000, 4200000, 300000],
        ['Healthcare', 'Medical Devices', 'Europe', 'Q1 2025', 2780000, 2900000, -120000],
        ['Healthcare', 'Pharmaceuticals', 'Asia Pacific', 'Q2 2025', 5200000, 5000000, 200000],
        ['Healthcare', 'Pharmaceuticals', 'North America', 'Q2 2025', 6100000, 5800000, 300000],
        ['Finance', 'Investment Banking', 'North America', 'Q1 2025', 8750000, 8500000, 250000],
        ['Finance', 'Investment Banking', 'Europe', 'Q1 2025', 4300000, 4500000, -200000],
        ['Finance', 'Retail Banking', 'North America', 'Q2 2025', 3650000, 3400000, 250000],
        ['Finance', 'Retail Banking', 'Asia Pacific', 'Q2 2025', 1980000, 2100000, -120000],
        ['Manufacturing', 'Automotive Parts', 'North America', 'Q1 2025', 7200000, 7000000, 200000],
        ['Manufacturing', 'Automotive Parts', 'Europe', 'Q1 2025', 5400000, 5500000, -100000],
        ['Manufacturing', 'Electronics', 'Asia Pacific', 'Q2 2025', 3100000, 3000000, 100000],
        ['Manufacturing', 'Electronics', 'North America', 'Q2 2025', 2850000, 2700000, 150000],
        ['Energy', 'Renewable', 'North America', 'Q1 2025', 4100000, 3800000, 300000],
        ['Energy', 'Renewable', 'Europe', 'Q1 2025', 3750000, 3600000, 150000],
        ['Energy', 'Oil & Gas', 'North America', 'Q2 2025', 9200000, 9000000, 200000],
        ['Energy', 'Oil & Gas', 'Asia Pacific', 'Q2 2025', 6800000, 7000000, -200000],
        ['Retail', 'E-Commerce', 'North America', 'Q1 2025', 5600000, 5300000, 300000],
        ['Retail', 'E-Commerce', 'Europe', 'Q2 2025', 3900000, 4000000, -100000],
        ['Retail', 'Brick & Mortar', 'North America', 'Q1 2025', 2400000, 2500000, -100000],
        ['Retail', 'Brick & Mortar', 'Asia Pacific', 'Q2 2025', 1750000, 1800000, -50000],
        ['Technology', 'AI Solutions', 'North America', 'Q2 2025', 4850000, 4500000, 350000],
        ['Healthcare', 'Telemedicine', 'Europe', 'Q2 2025', 1620000, 1500000, 120000],
        ['Finance', 'Wealth Management', 'North America', 'Q1 2025', 7300000, 7100000, 200000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
