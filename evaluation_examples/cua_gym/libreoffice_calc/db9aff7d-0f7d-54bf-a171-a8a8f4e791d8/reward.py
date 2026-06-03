"""
Reward Script: Archive Thunderbird 'Client Contracts' emails as .eml files,
               parse them with a Python script, save metadata to index.csv,
               and open in LibreOffice Calc sorted with AutoFilter enabled.
Task ID: osworld_multi_apps_email_file_convert_007
Domain: multi_apps (os + libreoffice_calc)
Scoring:
  Component 1: 6 .eml files in /home/user/contracts_backup/  (0.25 pts)
  Component 2: Python script at /home/user/scripts/parse_contracts_eml.py
               using the 'email' module                       (0.20 pts)
  Component 3: index.csv has 6 data rows with correct columns
               (subject, sender, date, attachments)            (0.25 pts)
  Component 4: index.xlsx has AutoFilter enabled               (0.20 pts)
  Component 5: index.xlsx data is sorted by date ascending     (0.10 pts)
  Total: 1.0
"""

import os
import csv

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_file_convert_007'
BACKUP_DIR = os.path.join(WORKDIR, 'contracts_backup')
SCRIPTS_DIR = os.path.join(WORKDIR, 'scripts')
SCRIPT_PATH = os.path.join(SCRIPTS_DIR, 'parse_contracts_eml.py')
INDEX_CSV = os.path.join(BACKUP_DIR, 'index.csv')
INDEX_XLSX = os.path.join(BACKUP_DIR, 'index.xlsx')


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # --- Precondition: backup directory exists ---
    if not os.path.isdir(BACKUP_DIR):
        print(f"CRITICAL: contracts_backup directory not found at {BACKUP_DIR}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 6 .eml files in /home/user/contracts_backup/ (0.25 pts)
    # Initial env has no contracts_backup dir, so all .eml checks are task-introduced changes
    try:
        all_files = os.listdir(BACKUP_DIR)
        eml_files = [f for f in all_files if f.endswith('.eml')]
        eml_count = len(eml_files)
        if eml_count == 6:
            print(f"PASS: Component 1 — Found {eml_count} .eml files in contracts_backup (0.25 pts)")
            total_score += 0.25
        elif eml_count > 0:
            # Partial: some eml files present
            partial = round(0.25 * (eml_count / 6), 4)
            print(f"PARTIAL: Component 1 — Found {eml_count}/6 .eml files; awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No .eml files found in {BACKUP_DIR}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Python script exists at /home/user/scripts/parse_contracts_eml.py
    #              and uses the 'email' module (0.20 pts)
    try:
        if not os.path.isfile(SCRIPT_PATH):
            print(f"FAIL: Component 2 — Script not found at {SCRIPT_PATH}")
        else:
            with open(SCRIPT_PATH, 'r', encoding='utf-8') as f:
                script_content = f.read()
            uses_email_module = 'import email' in script_content or 'from email' in script_content
            if uses_email_module:
                print(f"PASS: Component 2 — Script found and uses 'email' module (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — Script exists but does not use the 'email' module")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: index.csv has 6 data rows with correct columns (0.25 pts)
    try:
        if not os.path.isfile(INDEX_CSV):
            print(f"FAIL: Component 3 — index.csv not found at {INDEX_CSV}")
        else:
            with open(INDEX_CSV, 'r', encoding='utf-8', newline='') as f:
                reader = csv.DictReader(f)
                required_columns = {'subject', 'sender', 'date', 'attachments'}
                fieldnames = set(reader.fieldnames) if reader.fieldnames else set()
                has_all_columns = required_columns.issubset(fieldnames)
                rows = list(reader)
                row_count = len(rows)

            if has_all_columns and row_count == 6:
                print(f"PASS: Component 3 — index.csv has {row_count} data rows with correct columns "
                      f"(subject, sender, date, attachments) (0.25 pts)")
                total_score += 0.25
            elif has_all_columns and row_count > 0:
                partial = round(0.25 * (row_count / 6), 4)
                print(f"PARTIAL: Component 3 — index.csv has correct columns but {row_count}/6 rows; "
                      f"awarding {partial} pts")
                total_score += partial
            elif row_count == 6 and not has_all_columns:
                missing = required_columns - fieldnames
                print(f"FAIL: Component 3 — index.csv has 6 rows but missing columns: {missing}")
            else:
                print(f"FAIL: Component 3 — index.csv has {row_count} rows, "
                      f"columns={fieldnames}, expected 6 rows with {required_columns}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: index.xlsx has AutoFilter enabled (0.20 pts)
    try:
        if not os.path.isfile(INDEX_XLSX):
            print(f"FAIL: Component 4 — index.xlsx not found at {INDEX_XLSX}")
        else:
            import openpyxl
            wb = openpyxl.load_workbook(INDEX_XLSX)
            ws = wb.active
            auto_filter_ref = ws.auto_filter.ref
            if auto_filter_ref:
                print(f"PASS: Component 4 — index.xlsx has AutoFilter enabled "
                      f"(ref={auto_filter_ref}) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — index.xlsx does not have AutoFilter enabled")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: index.xlsx data is sorted by date ascending (0.10 pts)
    try:
        if not os.path.isfile(INDEX_XLSX):
            print(f"FAIL: Component 5 — index.xlsx not found, cannot check sorting")
        else:
            import openpyxl
            from email.utils import parsedate_to_datetime
            wb = openpyxl.load_workbook(INDEX_XLSX)
            ws = wb.active
            # Collect date values from column C (skip header row 1)
            date_values = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
                for cell in row:
                    val = cell.value
                    if val is not None:
                        date_values.append(str(val))

            # Attempt to parse all dates and check ascending order
            parsed_dates = []
            parse_errors = []
            for dv in date_values:
                try:
                    dt = parsedate_to_datetime(dv)
                    parsed_dates.append(dt)
                except Exception as pe:
                    parse_errors.append(f"{dv!r}: {pe}")

            if len(parse_errors) == 0 and len(parsed_dates) >= 2:
                # All dates parsed successfully — check ascending order
                is_sorted = all(parsed_dates[i] <= parsed_dates[i+1]
                                for i in range(len(parsed_dates)-1))
                if is_sorted:
                    print(f"PASS: Component 5 — index.xlsx data sorted by date ascending (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 5 — index.xlsx dates not in ascending order: {date_values}")
            elif len(parse_errors) > 0:
                print(f"FAIL: Component 5 — Could not parse date values: {parse_errors}")
            else:
                print(f"FAIL: Component 5 — Not enough date values found ({len(date_values)})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task()
