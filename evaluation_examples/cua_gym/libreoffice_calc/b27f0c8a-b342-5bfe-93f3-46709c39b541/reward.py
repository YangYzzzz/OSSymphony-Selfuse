"""
Reward Script: Extract PDF authors/institutions and build sorted LibreOffice Calc table
Task ID: osworld_multi_apps_pdf_author_extract_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Correct headers in row 1: 'Author Name', 'Institution', 'Paper Title'
  Component 2 (0.4): 8 data rows present with expected author names
  Component 3 (0.3): Data is sorted alphabetically by author name (A-Z)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_author_extract_007'

# Expected author names from the 8 PDFs (sorted order for reference)
EXPECTED_AUTHORS_SORTED = [
    'Aisha Mahmoud',
    'Brian Kowalski',
    'Chen Wei',
    'Diana Okonkwo',
    'Elena Vasquez',
    'Farhan Iqbal',
    'Grace Nakamura',
    'Henrik Sorensen',
]

EXPECTED_HEADERS = ['Author Name', 'Institution', 'Paper Title']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Use the first (active) sheet
    ws = wb.active

    # Component 1: Correct headers in row 1 (0.3 points)
    # Checks that the first row contains 'Author Name', 'Institution', 'Paper Title'
    try:
        header_row = [ws.cell(row=1, column=c).value for c in range(1, 4)]
        # Normalize: strip whitespace for comparison
        header_normalized = [str(h).strip() if h is not None else '' for h in header_row]
        if header_normalized == EXPECTED_HEADERS:
            print(f"PASS: Component 1 — Correct headers: {header_normalized} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Expected headers {EXPECTED_HEADERS}, found: {header_normalized}")
    except Exception as e:
        print(f"ERROR: Component 1 (headers) — {e}")

    # Component 2: 8 data rows with expected author names present (0.4 points)
    # Checks that all 8 expected author names appear somewhere in column A (rows 2-9)
    try:
        data_rows = []
        for r in range(2, ws.max_row + 1):
            name_val = ws.cell(row=r, column=1).value
            if name_val is not None:
                data_rows.append(str(name_val).strip())

        # Check row count
        if len(data_rows) != 8:
            print(f"FAIL: Component 2 — Expected 8 data rows, found: {len(data_rows)}")
        else:
            # Check all expected authors are present (case-insensitive match)
            data_lower = [n.lower() for n in data_rows]
            expected_lower = [n.lower() for n in EXPECTED_AUTHORS_SORTED]
            missing = [n for n in expected_lower if n not in data_lower]
            if not missing:
                print(f"PASS: Component 2 — All 8 expected author names present (0.4 pts)")
                total_score += 0.4
            else:
                print(f"FAIL: Component 2 — Missing authors: {missing}")
                print(f"      Found authors: {data_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 (author names) — {e}")

    # Component 3: Data sorted alphabetically by author name A-Z (0.3 points)
    # Checks that rows 2 onward are in ascending alphabetical order by column A values
    try:
        author_col = []
        for r in range(2, ws.max_row + 1):
            name_val = ws.cell(row=r, column=1).value
            if name_val is not None:
                author_col.append(str(name_val).strip())

        if len(author_col) >= 2:
            # Check that each name <= next name (case-insensitive)
            sorted_check = all(
                author_col[i].lower() <= author_col[i + 1].lower()
                for i in range(len(author_col) - 1)
            )
            if sorted_check:
                print(f"PASS: Component 3 — Data is sorted A-Z by author name (0.3 pts)")
                print(f"      Order: {author_col}")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — Data is NOT sorted A-Z by author name")
                print(f"      Found order: {author_col}")
                print(f"      Expected order: {sorted(author_col, key=str.lower)}")
        else:
            print(f"FAIL: Component 3 — Not enough rows to check sort order ({len(author_col)} rows)")
    except Exception as e:
        print(f"ERROR: Component 3 (sort order) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/hci_sys_authors.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
