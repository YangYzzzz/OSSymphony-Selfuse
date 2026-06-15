"""
Initial Setup: Create a packing slip template with raw data (no formulas)
Task ID: calc_ops_091
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_091'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PackSlip'

    # --- Title ---
    ws['A1'] = 'PACKING SLIP'
    ws['A1'].font = Font(name='Arial', size=16, bold=True)

    # --- Ship To / Order Info ---
    ws['A3'] = 'Ship To:'
    ws['A3'].font = Font(bold=True)
    ws['B3'] = 'Acme Corp'

    ws['A4'] = 'Order #:'
    ws['A4'].font = Font(bold=True)
    ws['B4'] = 'ORD-5001'

    # --- Column Headers (row 6) ---
    headers = ['Item', 'Description', 'Qty', 'Unit Price', 'Line Total']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Item Data (rows 7-10) - NO formulas in column E ---
    items = [
        ['ITM-01', 'Widget Pro', 10, 25.00],
        ['ITM-02', 'Widget Basic', 25, 12.50],
        ['ITM-03', 'Gadget X', 5, 45.00],
        ['ITM-04', 'Cable Kit', 50, 3.75],
    ]

    for r, row_data in enumerate(items, 7):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format Qty as integer, Unit Price as currency
    for r in range(7, 11):
        ws.cell(row=r, column=3).number_format = '0'
        ws.cell(row=r, column=4).number_format = '$#,##0.00'

    # --- Summary Labels (column A) ---
    ws['A12'] = 'Subtotal'
    ws['A12'].font = Font(bold=True)
    ws['A13'] = 'Tax (8%)'
    ws['A13'].font = Font(bold=True)
    ws['A14'] = 'Grand Total'
    ws['A14'].font = Font(bold=True, size=12)

    # Column E summary cells are intentionally left EMPTY (no formulas, no values)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # --- Thin borders on data area ---
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(6, 11):
        for c in range(1, 6):
            ws.cell(row=r, column=c).border = border

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
