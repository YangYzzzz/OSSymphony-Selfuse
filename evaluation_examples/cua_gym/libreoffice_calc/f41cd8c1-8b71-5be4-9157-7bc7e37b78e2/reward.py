"""
Reward Script: Packing slip template with auto-calculating formulas
Task ID: calc_ops_091
Domain: libreoffice_calc
Scoring:
  Component 1: Line total formulas in E7:E10 (=C*D for each row) — 0.4 pts (0.1 each)
  Component 2: Subtotal formula in E12 (=SUM(E7:E10)) — 0.2 pts
  Component 3: Tax formula in E13 (=E12*0.08) — 0.2 pts
  Component 4: Grand total formula in E14 (=E12+E13) — 0.2 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_091'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def is_multiply_formula(formula_str, row_num):
    """
    Check if formula multiplies C and D of the given row.
    Accepts patterns like =C7*D7, =D7*C7, etc.
    """
    f = normalize_formula(formula_str)
    if not f.startswith('='):
        return False
    # Expected patterns: =C{row}*D{row} or =D{row}*C{row}
    expected_a = f'=C{row_num}*D{row_num}'
    expected_b = f'=D{row_num}*C{row_num}'
    return f == expected_a or f == expected_b


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that PackSlip sheet exists (precondition gate)
    if 'PackSlip' not in wb.sheetnames:
        print(f"CRITICAL: 'PackSlip' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PackSlip']

    # Component 1: Line total formulas E7:E10 (0.4 points — 0.1 each)
    # Each cell should contain a formula multiplying Qty (C) by Unit Price (D)
    for row_num in [7, 8, 9, 10]:
        cell_ref = f'E{row_num}'
        try:
            val = ws[cell_ref].value
            if val is not None and is_multiply_formula(str(val), row_num):
                print(f"PASS: Component 1.{row_num-6} — {cell_ref} has multiply formula: {val} (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 1.{row_num-6} — {cell_ref} expected =C{row_num}*D{row_num}, found: {repr(val)}")
        except Exception as e:
            print(f"ERROR: Component 1.{row_num-6} — {cell_ref}: {e}")

    # Component 2: Subtotal formula in E12 (0.2 points)
    # Should be =SUM(E7:E10)
    try:
        val = ws['E12'].value
        f = normalize_formula(str(val)) if val else ''
        if f == '=SUM(E7:E10)':
            print(f"PASS: Component 2 — E12 has SUM formula: {val} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — E12 expected =SUM(E7:E10), found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — E12: {e}")

    # Component 3: Tax formula in E13 (0.2 points)
    # Should be =E12*0.08
    try:
        val = ws['E13'].value
        f = normalize_formula(str(val)) if val else ''
        if f == '=E12*0.08':
            print(f"PASS: Component 3 — E13 has tax formula: {val} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — E13 expected =E12*0.08, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — E13: {e}")

    # Component 4: Grand total formula in E14 (0.2 points)
    # Should be =E12+E13
    try:
        val = ws['E14'].value
        f = normalize_formula(str(val)) if val else ''
        if f == '=E12+E13':
            print(f"PASS: Component 4 — E14 has grand total formula: {val} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — E14 expected =E12+E13, found: {repr(val)}")
    except Exception as e:
        print(f"ERROR: Component 4 — E14: {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
