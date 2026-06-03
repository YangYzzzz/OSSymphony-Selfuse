"""
Initial Setup: Build a histogram chart from frequency data and add descriptive statistics alongside it.
Task ID: calc_gpm_025
Domain: libreoffice_calc

Creates the pre-task state: Sheet 'Analysis' with 40 numeric data points in column A.
No statistics, no frequency bins, no charts, no conditional formatting.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Analysis ---
    ws = wb.active
    ws.title = 'Analysis'

    # Column A header
    ws.cell(row=1, column=1, value='Value')

    # 40 numeric data points in A2:A41, roughly normally distributed around 55, range 10-100
    data_points = [
        52, 47, 63, 58, 71, 44, 55, 39, 67, 50,
        61, 48, 73, 56, 42, 65, 53, 78, 36, 59,
        45, 68, 54, 81, 40, 62, 49, 57, 70, 46,
        60, 33, 75, 51, 64, 43, 58, 85, 37, 55,
    ]

    for i, val in enumerate(data_points, start=2):
        ws.cell(row=i, column=1, value=val)

    # Set column A width for readability
    ws.column_dimensions['A'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
