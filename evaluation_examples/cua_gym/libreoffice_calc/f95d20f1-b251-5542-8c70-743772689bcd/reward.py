"""
Reward Script: Build a histogram chart from frequency data and add descriptive statistics
Task ID: calc_gpm_025
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Statistics labels C2:C10
  Component 2 (0.25): Statistics formulas D2:D10
  Component 3 (0.20): Frequency bins table F1:G9 with COUNTIFS
  Component 4 (0.15): Chart exists with correct type and title
  Component 5 (0.15): Conditional formatting on D6 (red) and D7 (green)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_025'


def persist_app_state(domain: str):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet 'Analysis' must exist
    if 'Analysis' not in wb.sheetnames:
        print("FAIL: Sheet 'Analysis' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Analysis']

    # ---------------------------------------------------------------
    # Component 1: Statistics labels in C2:C10 (0.25 points)
    # These labels do NOT exist in the initial file (all None).
    # ---------------------------------------------------------------
    try:
        expected_labels = {
            'C2': 'Count',
            'C3': 'Mean',
            'C4': 'Median',
            'C5': 'Std Dev',
            'C6': 'Min',
            'C7': 'Max',
            'C8': 'Range',
            'C9': 'Q1 (25th)',
            'C10': 'Q3 (75th)',
        }
        labels_found = 0
        for coord, expected in expected_labels.items():
            val = ws[coord].value
            if val is not None and str(val).strip().lower() == expected.lower():
                labels_found += 1
            else:
                print(f"  DETAIL: {coord} expected '{expected}', found '{val}'")

        if labels_found == 9:
            print(f"PASS: Component 1 -- All 9 statistics labels present (0.25 pts)")
            total_score += 0.25
        elif labels_found >= 5:
            partial = round(0.25 * labels_found / 9, 2)
            print(f"PARTIAL: Component 1 -- {labels_found}/9 labels found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {labels_found}/9 statistics labels found")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # ---------------------------------------------------------------
    # Component 2: Statistics formulas in D2:D10 (0.25 points)
    # These formulas do NOT exist in the initial file (all None).
    # ---------------------------------------------------------------
    try:
        expected_formulas = {
            'D2': '=COUNT(A2:A41)',
            'D3': '=AVERAGE(A2:A41)',
            'D4': '=MEDIAN(A2:A41)',
            'D5': '=STDEV(A2:A41)',
            'D6': '=MIN(A2:A41)',
            'D7': '=MAX(A2:A41)',
            'D8': '=D7-D6',
            'D9': '=PERCENTILE(A2:A41,0.25)',
            'D10': '=PERCENTILE(A2:A41,0.75)',
        }
        formulas_found = 0
        for coord, expected_formula in expected_formulas.items():
            val = ws[coord].value
            if val is not None and isinstance(val, str):
                # Normalize: uppercase, strip spaces
                norm_val = val.upper().replace(" ", "")
                norm_exp = expected_formula.upper().replace(" ", "")
                if norm_val == norm_exp:
                    formulas_found += 1
                else:
                    print(f"  DETAIL: {coord} expected '{expected_formula}', found '{val}'")
            elif val is not None:
                # Could be a computed numeric value (data_only mode or cached)
                # Still counts if it's a number (formula was evaluated)
                print(f"  DETAIL: {coord} has numeric value {val} (formula may have been evaluated)")
            else:
                print(f"  DETAIL: {coord} is None (no formula)")

        if formulas_found == 9:
            print(f"PASS: Component 2 -- All 9 statistics formulas correct (0.25 pts)")
            total_score += 0.25
        elif formulas_found >= 5:
            partial = round(0.25 * formulas_found / 9, 2)
            print(f"PARTIAL: Component 2 -- {formulas_found}/9 formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {formulas_found}/9 formulas correct")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # ---------------------------------------------------------------
    # Component 3: Frequency bins table F1:G9 with COUNTIFS (0.20 points)
    # F and G columns are empty in initial file.
    # ---------------------------------------------------------------
    try:
        # Check headers
        f1_val = ws['F1'].value
        g1_val = ws['G1'].value
        headers_ok = (f1_val is not None and 'bin' in str(f1_val).lower() and
                      g1_val is not None and 'count' in str(g1_val).lower())

        # Check bin labels exist in F2:F9
        bins_found = 0
        for r in range(2, 10):
            val = ws.cell(row=r, column=6).value
            if val is not None and str(val).strip():
                bins_found += 1

        # Check COUNTIFS formulas in G2:G9
        countifs_found = 0
        for r in range(2, 10):
            val = ws.cell(row=r, column=7).value
            if val is not None and isinstance(val, str) and 'COUNTIF' in val.upper():
                countifs_found += 1

        sub_score = 0.0
        if headers_ok:
            sub_score += 0.04
        if bins_found >= 7:
            sub_score += 0.06
        if countifs_found >= 7:
            sub_score += 0.10
        elif countifs_found >= 4:
            sub_score += 0.05

        if sub_score >= 0.18:
            print(f"PASS: Component 3 -- Frequency table complete: headers={headers_ok}, bins={bins_found}/8, countifs={countifs_found}/8 ({sub_score} pts)")
        elif sub_score > 0:
            print(f"PARTIAL: Component 3 -- headers={headers_ok}, bins={bins_found}/8, countifs={countifs_found}/8 ({sub_score} pts)")
        else:
            print(f"FAIL: Component 3 -- Frequency table missing or incomplete")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # ---------------------------------------------------------------
    # Component 4: Chart exists with correct type and title (0.15 points)
    # No charts exist in initial file.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]

            # Check chart type is column (col) or bar
            chart_type_ok = chart.type in ('col', 'bar')

            # Check chart title contains "histogram" (case-insensitive)
            title_text = ''
            if chart.title is not None:
                try:
                    # Try to extract title text from rich text
                    if hasattr(chart.title, 'tx') and chart.title.tx is not None:
                        rich = chart.title.tx.rich
                        if rich is not None:
                            for p in rich.paragraphs:
                                for run in p.r:
                                    title_text += run.t
                except Exception:
                    pass
                if not title_text:
                    title_text = str(chart.title)

            title_ok = 'histogram' in title_text.lower() or 'distribution' in title_text.lower()

            sub_score = 0.0
            if chart_type_ok:
                sub_score += 0.05
            if title_ok:
                sub_score += 0.05
            # Check gap width (should be 0 or very small for histogram style)
            gap_ok = False
            if hasattr(chart, 'gapWidth') and chart.gapWidth is not None:
                if float(chart.gapWidth) <= 10:
                    gap_ok = True
                    sub_score += 0.05
            if not gap_ok:
                print(f"  DETAIL: Gap width = {getattr(chart, 'gapWidth', 'N/A')}")

            if sub_score >= 0.14:
                print(f"PASS: Component 4 -- Chart: type={chart.type}, title='{title_text}', gapWidth={getattr(chart, 'gapWidth', 'N/A')} ({sub_score} pts)")
            elif sub_score > 0:
                print(f"PARTIAL: Component 4 -- type_ok={chart_type_ok}, title_ok={title_ok}, gap_ok={gap_ok} ({sub_score} pts)")
            else:
                print(f"FAIL: Component 4 -- Chart properties incorrect")
            total_score += sub_score
        else:
            print(f"FAIL: Component 4 -- No charts found")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # ---------------------------------------------------------------
    # Component 5: Conditional formatting on D6 (red) and D7 (green) (0.15 points)
    # No conditional formatting in initial file.
    # ---------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        d6_red = False
        d7_green = False

        for cf in cf_rules:
            cf_range = str(cf)
            for rule in cf.rules:
                if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                    fill_color = None
                    if rule.dxf.fill.fgColor:
                        fill_color = rule.dxf.fill.fgColor.rgb
                    elif rule.dxf.fill.bgColor:
                        fill_color = rule.dxf.fill.bgColor.rgb

                    if fill_color:
                        # Check for red on D6
                        if 'D6' in cf_range:
                            # Red: FFFF0000 or similar with high R, low G/B
                            r_val = int(fill_color[2:4], 16) if len(fill_color) >= 8 else 0
                            g_val = int(fill_color[4:6], 16) if len(fill_color) >= 8 else 0
                            if r_val > 200 and g_val < 100:
                                d6_red = True
                        # Check for green on D7
                        if 'D7' in cf_range:
                            g_val = int(fill_color[4:6], 16) if len(fill_color) >= 8 else 0
                            r_val = int(fill_color[2:4], 16) if len(fill_color) >= 8 else 0
                            if g_val > 200 and r_val < 100:
                                d7_green = True

        sub_score = 0.0
        if d6_red:
            sub_score += 0.075
        if d7_green:
            sub_score += 0.075

        if sub_score >= 0.14:
            print(f"PASS: Component 5 -- CF: D6 red={d6_red}, D7 green={d7_green} ({sub_score} pts)")
        elif sub_score > 0:
            print(f"PARTIAL: Component 5 -- D6 red={d6_red}, D7 green={d7_green} ({sub_score} pts)")
        else:
            print(f"FAIL: Component 5 -- No conditional formatting for D6/D7")
        total_score += sub_score
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
