"""
Reward Script: Set up a formatted attendance register with date headers and present/absent color coding.
Task ID: calc_gpm_039
Domain: libreoffice_calc
Scoring:
  Component 1: Title row - merged A1:L1, correct text, bold 13pt (0.15)
  Component 2: Date headers in row 2 with vertical text orientation (0.15)
  Component 3: Student names in A3:A17 - 15 students (0.10)
  Component 4: Attendance data P/A in B3:L17 grid (0.10)
  Component 5: Conditional formatting on B3:L17 for P/A color coding (0.15)
  Component 6: COUNTIF formulas in M/N/O columns (0.15)
  Component 7: Daily attendance formulas in row 19 (0.10)
  Component 8: Column widths and borders (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_039'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        import time
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Attendance' sheet must exist
    if 'Attendance' not in wb.sheetnames:
        print("FAIL: 'Attendance' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Attendance']

    # Component 1: Title row - merged A1:L1 with correct text, bold, 13pt (0.15 points)
    try:
        comp1 = 0.0
        # Check merge
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'L1' in r for r in merged_ranges)
        if has_merge:
            comp1 += 0.05

        # Check title text
        title_val = ws['A1'].value
        if title_val and 'Attendance Register' in str(title_val) and 'April 2026' in str(title_val):
            comp1 += 0.05

        # Check bold and size 13
        font = ws['A1'].font
        if font.bold and font.size and abs(font.size - 13) < 0.5:
            comp1 += 0.05

        if comp1 > 0:
            print(f"PASS: Component 1 - Title row ({comp1:.2f} pts) merge={has_merge}, text='{title_val}', bold={font.bold}, size={font.size}")
            total_score += comp1
        else:
            print(f"FAIL: Component 1 - Title row: value={title_val}, bold={font.bold}, size={font.size}, merges={merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Date headers in row 2 with vertical text orientation (0.15 points)
    try:
        comp2 = 0.0
        # Check B2:L2 have date values
        date_count = 0
        for col in range(2, 13):  # B=2 to L=12
            val = ws.cell(row=2, column=col).value
            if val is not None and str(val).strip():
                date_count += 1

        if date_count >= 9:  # 11 weekdays expected, allow some tolerance
            comp2 += 0.05
            print(f"  Date headers found: {date_count}/11")

        # Check vertical text orientation (textRotation=90 or similar)
        rotation_count = 0
        for col in range(2, 13):
            cell = ws.cell(row=2, column=col)
            if cell.alignment and cell.alignment.textRotation and cell.alignment.textRotation > 0:
                rotation_count += 1

        if rotation_count >= 9:
            comp2 += 0.05
            print(f"  Vertical rotation count: {rotation_count}/11")

        # Check bold
        bold_count = 0
        for col in range(2, 13):
            if ws.cell(row=2, column=col).font.bold:
                bold_count += 1
        if bold_count >= 9:
            comp2 += 0.05
            print(f"  Bold headers: {bold_count}/11")

        if comp2 > 0:
            print(f"PASS: Component 2 - Date headers ({comp2:.2f} pts)")
            total_score += comp2
        else:
            print(f"FAIL: Component 2 - Date headers: dates={date_count}, rotations={rotation_count}, bold={bold_count}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Student names in A3:A17 - 15 students (0.10 points)
    try:
        student_count = 0
        for row in range(3, 18):
            val = ws.cell(row=row, column=1).value
            if val is not None and str(val).strip():
                student_count += 1

        if student_count >= 14:
            print(f"PASS: Component 3 - {student_count} students found (0.10 pts)")
            total_score += 0.10
        elif student_count >= 10:
            pts = 0.05
            print(f"PARTIAL: Component 3 - {student_count}/15 students ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 3 - Only {student_count} students found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Attendance data P/A in B3:L17 grid (0.10 points)
    try:
        pa_count = 0
        total_cells = 0
        for row in range(3, 18):
            for col in range(2, 13):
                val = ws.cell(row=row, column=col).value
                total_cells += 1
                if val is not None and str(val).strip().upper() in ('P', 'A'):
                    pa_count += 1

        ratio = pa_count / total_cells if total_cells > 0 else 0
        if ratio >= 0.90:
            print(f"PASS: Component 4 - Attendance data: {pa_count}/{total_cells} cells are P/A (0.10 pts)")
            total_score += 0.10
        elif ratio >= 0.50:
            pts = 0.05
            print(f"PARTIAL: Component 4 - {pa_count}/{total_cells} P/A cells ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 4 - Only {pa_count}/{total_cells} P/A cells")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting on B3:L17 for P/A color coding (0.15 points)
    try:
        comp5 = 0.0
        pa_cf_count = 0
        rate_cf_count = 0

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                # Check for P/A conditional formatting on attendance area
                if 'B3' in cf_range and 'L17' in cf_range:
                    if hasattr(rule, 'formula') and rule.formula:
                        for f in rule.formula:
                            if '"P"' in str(f) or '"A"' in str(f):
                                pa_cf_count += 1
                    if rule.type == 'cellIs':
                        pa_cf_count += 1

                # Check for rate conditional formatting on O column
                if 'O3' in cf_range or 'O17' in cf_range:
                    rate_cf_count += 1

        if pa_cf_count > 0:
            comp5 += 0.10
            print(f"  P/A conditional formatting found on B3:L17 ({pa_cf_count} rules)")
        if rate_cf_count > 0:
            comp5 += 0.05
            print(f"  Rate conditional formatting found on O column ({rate_cf_count} rules)")

        if comp5 > 0:
            print(f"PASS: Component 5 - Conditional formatting ({comp5:.2f} pts)")
            total_score += comp5
        else:
            print(f"FAIL: Component 5 - No relevant conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: COUNTIF formulas in M/N/O columns (0.15 points)
    try:
        comp6 = 0.0
        # Check M column - COUNTIF for P
        m_formula_count = 0
        for row in range(3, 18):
            val = ws.cell(row=row, column=13).value  # M column
            if val and isinstance(val, str) and 'COUNTIF' in val.upper() and '"P"' in val:
                m_formula_count += 1

        if m_formula_count >= 13:
            comp6 += 0.05
            print(f"  M column COUNTIF('P') formulas: {m_formula_count}/15")

        # Check N column - COUNTIF for A
        n_formula_count = 0
        for row in range(3, 18):
            val = ws.cell(row=row, column=14).value  # N column
            if val and isinstance(val, str) and 'COUNTIF' in val.upper() and '"A"' in val:
                n_formula_count += 1

        if n_formula_count >= 13:
            comp6 += 0.05
            print(f"  N column COUNTIF('A') formulas: {n_formula_count}/15")

        # Check O column - rate formula (M/(M+N) pattern)
        o_formula_count = 0
        for row in range(3, 18):
            val = ws.cell(row=row, column=15).value  # O column
            if val and isinstance(val, str) and ('M' in val.upper() or 'N' in val.upper()):
                # Check it's some kind of ratio formula
                o_formula_count += 1

        if o_formula_count >= 13:
            comp6 += 0.05
            print(f"  O column rate formulas: {o_formula_count}/15")

        if comp6 > 0:
            print(f"PASS: Component 6 - COUNTIF formulas ({comp6:.2f} pts)")
            total_score += comp6
        else:
            print(f"FAIL: Component 6 - M formulas={m_formula_count}, N formulas={n_formula_count}, O formulas={o_formula_count}")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Daily attendance formulas in row 19 (0.10 points)
    try:
        comp7 = 0.0
        # Check A19 has 'Daily Attendance' label and is bold
        a19 = ws['A19'].value
        a19_bold = ws['A19'].font.bold

        if a19 and 'Daily' in str(a19) and 'Attendance' in str(a19) and a19_bold:
            comp7 += 0.04
            print(f"  A19: '{a19}', bold={a19_bold}")

        # Check B19:L19 have COUNTIF formulas
        daily_formula_count = 0
        for col in range(2, 13):
            val = ws.cell(row=19, column=col).value
            if val and isinstance(val, str) and 'COUNTIF' in val.upper():
                daily_formula_count += 1

        if daily_formula_count >= 9:
            comp7 += 0.06
            print(f"  Daily formulas found: {daily_formula_count}/11")

        if comp7 > 0:
            print(f"PASS: Component 7 - Daily attendance row ({comp7:.2f} pts)")
            total_score += comp7
        else:
            print(f"FAIL: Component 7 - A19={a19}, bold={a19_bold}, daily formulas={daily_formula_count}")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    # Component 8: Column widths and borders (0.10 points)
    try:
        comp8 = 0.0

        # Check column A width ~20
        col_a_width = ws.column_dimensions['A'].width
        if col_a_width and abs(col_a_width - 20) <= 2:
            comp8 += 0.03
            print(f"  Column A width: {col_a_width}")

        # Check column B width ~4
        col_b_width = ws.column_dimensions['B'].width
        if col_b_width and abs(col_b_width - 4) <= 1.5:
            comp8 += 0.02
            print(f"  Column B width: {col_b_width}")

        # Check borders on attendance area - sample a few cells
        border_count = 0
        sample_cells = ['B3', 'F5', 'L10', 'A3', 'M3']
        for coord in sample_cells:
            cell = ws[coord]
            if (cell.border.left.style and cell.border.right.style and
                    cell.border.top.style and cell.border.bottom.style):
                border_count += 1

        if border_count >= 3:
            comp8 += 0.05
            print(f"  Borders found on {border_count}/{len(sample_cells)} sample cells")

        if comp8 > 0:
            print(f"PASS: Component 8 - Column widths & borders ({comp8:.2f} pts)")
            total_score += comp8
        else:
            print(f"FAIL: Component 8 - colA={col_a_width}, colB={col_b_width}, borders={border_count}")
    except Exception as e:
        print(f"ERROR: Component 8 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
