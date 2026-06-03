"""
Initial Setup: Product comparison matrix with 8 products and 12 features
Task ID: calc_wf_039
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Comparison"

    # --- Layout ---
    # Row 1: Headers
    # Col A: Feature name
    # Col B: Weight (%)
    # Cols C-J: 8 products (scores 1-5)

    products = [
        "CloudSync Pro", "DataVault X", "NetGuard Elite",
        "StreamLine 360", "PixelForge AI", "CodeNexus Hub",
        "SecureWave Plus", "AnalytiCore V2"
    ]

    features = [
        ("Price Competitiveness", 0.12),
        ("Build Quality", 0.10),
        ("Customer Support", 0.08),
        ("Ease of Use", 0.10),
        ("Scalability", 0.09),
        ("Security Features", 0.11),
        ("Integration Options", 0.08),
        ("Performance Speed", 0.10),
        ("Documentation", 0.06),
        ("Mobile Compatibility", 0.07),
        ("Customization Depth", 0.05),
        ("Reliability / Uptime", 0.04),
    ]

    # Scores matrix (12 features x 8 products), values 1-5
    scores = [
        [4, 3, 5, 4, 2, 3, 5, 4],   # Price Competitiveness
        [3, 5, 4, 3, 4, 5, 3, 4],   # Build Quality
        [5, 4, 3, 4, 3, 4, 5, 3],   # Customer Support
        [4, 3, 4, 5, 3, 2, 4, 4],   # Ease of Use
        [3, 5, 4, 4, 5, 4, 3, 5],   # Scalability
        [4, 4, 5, 3, 3, 5, 5, 4],   # Security Features
        [3, 4, 3, 5, 4, 4, 3, 4],   # Integration Options
        [5, 4, 3, 4, 5, 3, 4, 5],   # Performance Speed
        [3, 3, 4, 4, 3, 5, 3, 4],   # Documentation
        [4, 2, 3, 5, 4, 3, 4, 3],   # Mobile Compatibility
        [2, 4, 3, 4, 5, 4, 2, 3],   # Customization Depth
        [4, 5, 4, 3, 4, 4, 5, 4],   # Reliability / Uptime
    ]

    # --- Row 1: Headers ---
    ws.cell(row=1, column=1, value="Feature")
    ws.cell(row=1, column=2, value="Weight")
    for i, prod in enumerate(products):
        ws.cell(row=1, column=3 + i, value=prod)

    # Basic header styling (plain bold, no colors - task asks agent to add color-coded headers)
    header_font = Font(bold=True, size=11)
    for col in range(1, 11):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # --- Rows 2-13: Features with weights and scores ---
    for r, (feature_name, weight) in enumerate(features, 2):
        ws.cell(row=r, column=1, value=feature_name)
        weight_cell = ws.cell(row=r, column=2, value=weight)
        weight_cell.number_format = '0%'
        weight_cell.alignment = Alignment(horizontal="center")

        for c, score in enumerate(scores[r - 2], 3):
            score_cell = ws.cell(row=r, column=c, value=score)
            score_cell.alignment = Alignment(horizontal="center")

    # --- Row 14: Label row for weighted totals (empty - agent fills) ---
    ws.cell(row=14, column=1, value="Weighted Total")
    ws.cell(row=14, column=1).font = Font(bold=True)

    # --- Row 15: Label row for ranking (empty - agent fills) ---
    ws.cell(row=15, column=1, value="Rank")
    ws.cell(row=15, column=1).font = Font(bold=True)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 10
    for col_letter in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 16

    # Row height for header
    ws.row_dimensions[1].height = 35

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
