"""
Reward Script: Product Comparison Matrix
Task ID: calc_wf_039
Domain: libreoffice_calc
Scoring:
  Component 1: SUMPRODUCT formulas in row 14 (0.25 pts)
  Component 2: RANK formulas in row 15 (0.20 pts)
  Component 3: Icon set conditional formatting on scores (0.15 pts)
  Component 4: Radar chart with 3 series (0.20 pts)
  Component 5: Color-coded column headers (gold/silver/bronze) (0.10 pts)
  Component 6: Alternating row shading on feature rows (0.10 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_039'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Comparison' must exist
    if 'Comparison' not in wb.sheetnames:
        print("FAIL: Sheet 'Comparison' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Comparison']

    # Component 1: SUMPRODUCT formulas in row 14, columns C-J (0.25 points)
    # Initial env has empty cells in row 14; golden has SUMPRODUCT formulas
    try:
        sumproduct_count = 0
        for col in range(3, 11):  # C=3 to J=10
            val = ws.cell(row=14, column=col).value
            if val is not None and isinstance(val, str) and 'SUMPRODUCT' in val.upper():
                sumproduct_count += 1
        if sumproduct_count == 8:
            print(f"PASS: Component 1 — All 8 SUMPRODUCT formulas found in row 14 (0.25 pts)")
            total_score += 0.25
        elif sumproduct_count >= 4:
            partial = round(0.25 * sumproduct_count / 8, 2)
            print(f"PARTIAL: Component 1 — {sumproduct_count}/8 SUMPRODUCT formulas in row 14 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Expected SUMPRODUCT formulas in row 14, found {sumproduct_count}/8")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: RANK formulas in row 15, columns C-J (0.20 points)
    # Initial env has empty cells in row 15; golden has RANK formulas
    try:
        rank_count = 0
        for col in range(3, 11):  # C=3 to J=10
            val = ws.cell(row=15, column=col).value
            if val is not None and isinstance(val, str) and 'RANK' in val.upper():
                rank_count += 1
        if rank_count == 8:
            print(f"PASS: Component 2 — All 8 RANK formulas found in row 15 (0.20 pts)")
            total_score += 0.20
        elif rank_count >= 4:
            partial = round(0.20 * rank_count / 8, 2)
            print(f"PARTIAL: Component 2 — {rank_count}/8 RANK formulas in row 15 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected RANK formulas in row 15, found {rank_count}/8")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Icon set conditional formatting on score cells C2:J13 (0.15 points)
    # Initial env has no conditional formatting; golden has iconSet rule
    try:
        icon_set_found = False
        cf_list = list(ws.conditional_formatting)
        for cfr in cf_list:
            for rule in cfr.rules:
                if rule.type == 'iconSet':
                    icon_set_found = True
                    break
            if icon_set_found:
                break

        if icon_set_found:
            print(f"PASS: Component 3 — Icon set conditional formatting found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — No icon set conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Radar chart with 3 series for top 3 products (0.20 points)
    # Initial env has no charts; golden has a radar chart with 3 series
    try:
        charts = ws._charts
        radar_found = False
        radar_series_count = 0
        radar_title_ok = False

        for ch in charts:
            chart_type = type(ch).__name__
            if 'Radar' in chart_type:
                radar_found = True
                radar_series_count = len(ch.series)
                # Check title contains something about top 3 or comparison
                if ch.title is not None:
                    try:
                        # title can be a Title object with nested text
                        title_text = ''
                        if hasattr(ch.title, 'text'):
                            title_text = str(ch.title.text) if ch.title.text else ''
                        if not title_text and hasattr(ch.title, 'tx') and ch.title.tx:
                            if hasattr(ch.title.tx, 'rich') and ch.title.tx.rich:
                                for p in ch.title.tx.rich.p:
                                    for r in p.r:
                                        title_text += r.t
                        if title_text:
                            radar_title_ok = True
                    except Exception:
                        radar_title_ok = True  # title exists but hard to extract
                break

        if radar_found and radar_series_count == 3:
            print(f"PASS: Component 4 — Radar chart with 3 series found (0.20 pts)")
            total_score += 0.20
        elif radar_found and radar_series_count > 0:
            partial = round(0.20 * min(radar_series_count, 3) / 3, 2)
            print(f"PARTIAL: Component 4 — Radar chart found with {radar_series_count} series ({partial} pts)")
            total_score += partial
        elif len(charts) > 0:
            # Some chart exists but not radar type
            print(f"PARTIAL: Component 4 — Chart found but not radar type (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Color-coded column headers for top 3 products (0.10 points)
    # Initial env has no fill colors on headers; golden has gold/silver/bronze
    # Gold = FFD700, Silver = C0C0C0, Bronze = CD7F32
    try:
        gold_found = False
        silver_found = False
        bronze_found = False

        for col in range(3, 11):
            cell = ws.cell(row=1, column=col)
            try:
                fill_rgb = cell.fill.fgColor.rgb
                fill_type = cell.fill.fill_type
                if fill_type != 'solid':
                    continue
            except Exception:
                continue

            if fill_rgb is None:
                continue

            # Normalize: check last 6 chars (strip alpha)
            color_6 = fill_rgb[-6:].upper()

            if color_6 == 'FFD700':
                gold_found = True
            elif color_6 == 'C0C0C0':
                silver_found = True
            elif color_6 == 'CD7F32':
                bronze_found = True

        colors_found = sum([gold_found, silver_found, bronze_found])
        if colors_found == 3:
            print(f"PASS: Component 5 — Gold, silver, bronze header colors found (0.10 pts)")
            total_score += 0.10
        elif colors_found > 0:
            partial = round(0.10 * colors_found / 3, 2)
            print(f"PARTIAL: Component 5 — {colors_found}/3 color-coded headers found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — No color-coded headers found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Alternating row shading on feature rows (0.10 points)
    # Initial env has no row shading; golden has FFF2F2F2 on even rows (2,4,6,8,10,12)
    try:
        shaded_even_rows = 0
        unshaded_odd_rows = 0
        even_rows = [2, 4, 6, 8, 10, 12]
        odd_rows = [3, 5, 7, 9, 11, 13]

        for r in even_rows:
            cell = ws.cell(row=r, column=1)
            try:
                fill_rgb = cell.fill.fgColor.rgb
                fill_type = cell.fill.fill_type
                if fill_type == 'solid' and fill_rgb and fill_rgb != '00000000':
                    shaded_even_rows += 1
            except Exception:
                pass

        for r in odd_rows:
            cell = ws.cell(row=r, column=1)
            try:
                fill_rgb = cell.fill.fgColor.rgb
                fill_type = cell.fill.fill_type
                # Odd rows should NOT have shading (or have different shading)
                if fill_type != 'solid' or fill_rgb in (None, '00000000'):
                    unshaded_odd_rows += 1
            except Exception:
                unshaded_odd_rows += 1  # no fill = unshaded

        # Need alternating pattern: shaded even, unshaded odd
        if shaded_even_rows >= 5 and unshaded_odd_rows >= 5:
            print(f"PASS: Component 6 — Alternating row shading found ({shaded_even_rows}/6 even shaded, {unshaded_odd_rows}/6 odd unshaded) (0.10 pts)")
            total_score += 0.10
        elif shaded_even_rows >= 3:
            partial = round(0.10 * shaded_even_rows / 6, 2)
            print(f"PARTIAL: Component 6 — Partial alternating shading ({shaded_even_rows}/6 even rows shaded) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — No alternating row shading detected (even rows shaded: {shaded_even_rows}/6)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
