"""
Initial Setup: Meeting minutes template with action items tracker
Task ID: calc_wf_067
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Meeting"

    # --- Styling helpers ---
    header_font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
    label_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    table_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    table_header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # --- Column widths ---
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14

    # --- Header Section (Rows 1-6) ---
    ws.merge_cells("A1:E1")
    ws["A1"] = "Meeting Minutes"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws["A2"] = "Date:"
    ws["A2"].font = label_font
    ws["B2"] = "March 18, 2025"
    ws["B2"].font = normal_font

    ws["A3"] = "Location:"
    ws["A3"].font = label_font
    ws["B3"] = "Conference Room B - Building 3"
    ws["B3"].font = normal_font

    ws["A4"] = "Attendees:"
    ws["A4"].font = label_font
    ws["B4"] = "Sarah Chen, Marcus Johnson, Priya Patel, David Kim, Rachel Torres"
    ws["B4"].font = normal_font

    ws["A5"] = "Purpose:"
    ws["A5"].font = label_font
    ws["B5"] = "Q1 Sprint Retrospective and Q2 Planning"
    ws["B5"].font = normal_font

    ws["A6"] = "Facilitator:"
    ws["A6"].font = label_font
    ws["B6"] = "Sarah Chen"
    ws["B6"].font = normal_font

    # --- Agenda Section (Rows 8-15) ---
    ws.merge_cells("A8:E8")
    ws["A8"] = "AGENDA"
    ws["A8"].font = header_font
    ws["A8"].alignment = Alignment(horizontal="left")

    agenda_items = [
        "1. Review of Q1 sprint deliverables and velocity metrics",
        "2. Team feedback on process improvements",
        "3. Discussion of blockers and technical debt",
        "4. Q2 roadmap priorities and resource allocation",
        "5. Customer feedback integration plan",
        "6. Update on cloud migration timeline",
        "7. Next steps and action item assignment",
    ]
    for i, item in enumerate(agenda_items):
        row = 9 + i
        ws.merge_cells(f"A{row}:E{row}")
        ws[f"A{row}"] = item
        ws[f"A{row}"].font = normal_font

    # --- Action Items Table (Row 17+) ---
    ws.merge_cells("A17:E17")
    ws["A17"] = "ACTION ITEMS"
    ws["A17"].font = header_font

    # Table headers in row 18
    table_headers = ["Item", "Owner", "Due Date", "Priority", "Status"]
    for col, h in enumerate(table_headers, 1):
        cell = ws.cell(row=18, column=col, value=h)
        cell.font = table_header_font
        cell.fill = table_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Action item data (rows 19-28)
    today = date(2025, 3, 18)
    action_items = [
        ["Compile Q1 velocity report and share with stakeholders", "Marcus Johnson", date(2025, 3, 21), "H", "Open"],
        ["Update CI/CD pipeline configuration for new microservices", "David Kim", date(2025, 3, 25), "H", "In Progress"],
        ["Draft Q2 product roadmap document", "Sarah Chen", date(2025, 3, 28), "H", "Open"],
        ["Schedule customer feedback sessions for April", "Rachel Torres", date(2025, 3, 20), "M", "Done"],
        ["Resolve outstanding security audit findings", "David Kim", date(2025, 3, 14), "H", "Open"],
        ["Migrate staging environment to Kubernetes cluster", "Marcus Johnson", date(2025, 4, 4), "M", "In Progress"],
        ["Prepare sprint retrospective summary for leadership", "Priya Patel", date(2025, 3, 19), "M", "Done"],
        ["Review and update API documentation for v2.3", "Priya Patel", date(2025, 3, 26), "L", "Open"],
        ["Set up monitoring dashboards for new services", "David Kim", date(2025, 4, 1), "M", "Open"],
        ["Coordinate with design team on Q2 UI refresh", "Rachel Torres", date(2025, 4, 7), "L", "Open"],
    ]

    for r, row_data in enumerate(action_items, 19):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = normal_font
            cell.border = thin_border
            if c == 3 and isinstance(val, date):
                cell.number_format = 'yyyy-mm-dd'
            if c in (2, 4, 5):
                cell.alignment = Alignment(horizontal="center")

    # --- Summary Section (Row 30+) ---
    ws.merge_cells("A30:E30")
    ws["A30"] = "SUMMARY"
    ws["A30"].font = header_font

    # Labels only - NO formulas (those are the task)
    ws["A31"] = "Open Items:"
    ws["A31"].font = label_font
    ws["A32"] = "Completed Items:"
    ws["A32"].font = label_font
    ws["A33"] = "Overdue Items:"
    ws["A33"].font = label_font
    ws["A34"] = "Total Items:"
    ws["A34"].font = label_font
    ws["B34"] = len(action_items)
    ws["B34"].font = normal_font

    # B31, B32, B33 left empty - task requires adding formulas here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
