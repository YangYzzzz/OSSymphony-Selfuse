"""
Reward Script: Meeting minutes template with action items tracker
Task ID: calc_wf_067
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30) - Summary formulas (B31 COUNTIF Open+InProgress, B32 COUNTIF Done, B33 COUNTIFS overdue)
  Component 2 (0.30) - Conditional formatting (3 rules: Done=green, overdue=red, H=red font)
  Component 3 (0.20) - Print area set to A1:E34
  Component 4 (0.10) - Header/Footer configured
  Component 5 (0.10) - Page break at row 16 (between agenda and action items)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_067'


def check_formula_cell(ws, coord, required_keywords):
    """Check if cell contains a formula with all required keywords (case-insensitive, space-stripped)."""
    val = ws[coord].value
    if val and isinstance(val, str):
        norm = val.upper().replace(" ", "")
        if all(kw in norm for kw in required_keywords):
            return val
    return None


def check_cf_done_green(cf_rules):
    """Check for conditional formatting rule: Status 'Done' -> green fill."""
    for cf in cf_rules:
        for rule in cf.rules:
            if rule.type == 'cellIs' and getattr(rule, 'operator', None) == 'equal':
                for f in (rule.formula or []):
                    if 'Done' in f or 'done' in f or 'DONE' in f:
                        if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                            rgb_hex = rule.dxf.fill.fgColor.rgb or ''
                            if '92D050' in rgb_hex or '00FF00' in rgb_hex or '00B050' in rgb_hex:
                                return rgb_hex
    return None


def check_cf_overdue_red(cf_rules):
    """Check for conditional formatting rule: overdue items -> red fill."""
    for cf in cf_rules:
        for rule in cf.rules:
            if rule.type == 'expression':
                for f in (rule.formula or []):
                    f_up = f.upper().replace(" ", "")
                    if 'TODAY' in f_up and 'DONE' in f_up:
                        if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                            rgb_hex = rule.dxf.fill.fgColor.rgb or ''
                            if 'FF0000' in rgb_hex:
                                return rgb_hex
    return None


def check_cf_priority_h_red(cf_rules):
    """Check for conditional formatting rule: Priority 'H' -> red font."""
    for cf in cf_rules:
        for rule in cf.rules:
            if rule.type == 'cellIs' and getattr(rule, 'operator', None) == 'equal':
                for f in (rule.formula or []):
                    if f.strip('"').strip("'") == 'H':
                        if rule.dxf and rule.dxf.font and rule.dxf.font.color:
                            fc = rule.dxf.font.color.rgb or ''
                            if 'FF0000' in fc:
                                return fc
    return None


def check_header_footer_text(hf_obj):
    """Check if a header or footer object has any non-empty text."""
    if hf_obj:
        for part in [hf_obj.left, hf_obj.center, hf_obj.right]:
            if part and part.text and len(part.text.strip()) > 0:
                return part.text.strip()
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Meeting' sheet must exist
    if 'Meeting' not in wb.sheetnames:
        print("FAIL: 'Meeting' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Meeting']

    # =========================================================================
    # Component 1: Summary formulas in B31, B32, B33 (0.30 points)
    # These cells are empty (None) in initial_env, formulas in golden_env
    # =========================================================================
    try:
        # B31: Open count = COUNTIF(status,"Open") + COUNTIF(status,"In Progress")
        b31_val = check_formula_cell(ws, 'B31', ['COUNTIF', 'OPEN', 'INPROGRESS'])
        if b31_val:
            print(f"PASS: B31 has open items formula: {b31_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: B31 missing or wrong formula. Found: {ws['B31'].value}")

        # B32: Done count = COUNTIF(status,"Done")
        b32_val = check_formula_cell(ws, 'B32', ['COUNTIF', 'DONE'])
        if b32_val:
            print(f"PASS: B32 has done count formula: {b32_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: B32 missing or wrong formula. Found: {ws['B32'].value}")

        # B33: Overdue = COUNTIFS(status<>"Done", due<TODAY())
        b33_val = check_formula_cell(ws, 'B33', ['COUNTIF', 'TODAY'])
        if b33_val:
            print(f"PASS: B33 has overdue formula: {b33_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: B33 missing or wrong formula. Found: {ws['B33'].value}")

    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Conditional formatting rules (0.30 points)
    # Initial has 0 rules, golden has 3 rules
    # =========================================================================
    try:
        cf_rules = list(ws.conditional_formatting)

        if len(cf_rules) == 0:
            print("FAIL: No conditional formatting rules found")
        else:
            # Sub-check 2a: Done -> green fill
            done_color = check_cf_done_green(cf_rules)
            if done_color:
                print(f"PASS: CF Done=green fill ({done_color}) (0.10 pts)")
                total_score += 0.10
            else:
                print("FAIL: No CF rule for Done=green fill found")

            # Sub-check 2b: Overdue -> red fill
            overdue_color = check_cf_overdue_red(cf_rules)
            if overdue_color:
                print(f"PASS: CF overdue=red fill ({overdue_color}) (0.10 pts)")
                total_score += 0.10
            else:
                print("FAIL: No CF rule for overdue=red fill found")

            # Sub-check 2c: Priority H -> red font
            priority_color = check_cf_priority_h_red(cf_rules)
            if priority_color:
                print(f"PASS: CF Priority H=red font ({priority_color}) (0.10 pts)")
                total_score += 0.10
            else:
                print("FAIL: No CF rule for Priority H=red font found")

    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Print area set (0.20 points)
    # Initial has no print area, golden has 'Meeting'!$A$1:$E$34
    # =========================================================================
    try:
        print_area = ws.print_area
        if print_area and len(str(print_area)) > 0:
            pa_str = str(print_area)
            # Check that it covers columns A through E
            if 'A' in pa_str and 'E' in pa_str:
                print(f"PASS: Print area is set: {pa_str} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Print area set but doesn't cover expected columns A-E. Found: {pa_str}")
        else:
            print("FAIL: No print area set")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =========================================================================
    # Component 4: Header/Footer configured (0.10 points)
    # Initial has no header/footer, golden has header center + footer left/right
    # =========================================================================
    try:
        header_text = check_header_footer_text(ws.oddHeader)
        footer_text = check_header_footer_text(ws.oddFooter)

        if header_text and footer_text:
            print(f"PASS: Both header and footer configured (0.10 pts)")
            total_score += 0.10
        elif header_text or footer_text:
            which = 'header' if header_text else 'footer'
            print(f"PARTIAL: Only {which} configured (0.05 pts)")
            total_score += 0.05
        else:
            print("FAIL: No header or footer configured")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =========================================================================
    # Component 5: Page break set (0.10 points)
    # Initial has no page breaks, golden has row break at 16
    # =========================================================================
    try:
        row_breaks = ws.row_breaks.brk if ws.row_breaks else []
        if len(row_breaks) > 0:
            break_ids = [brk.id for brk in row_breaks]
            print(f"PASS: Page break(s) set at row(s) {break_ids} (0.10 pts)")
            total_score += 0.10
        else:
            print("FAIL: No page breaks set")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
