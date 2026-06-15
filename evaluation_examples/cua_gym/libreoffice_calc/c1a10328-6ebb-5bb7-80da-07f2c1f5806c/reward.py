"""
Reward Script: Build out HR interview scorecard with AVERAGE scores, RANK, and Recommend column
Task ID: calc_hr_interview_scorecard_022
Domain: libreoffice_calc
Scoring:
  Component 1: H2:H34 contain =AVERAGE(Cx:Gx) formula        — 0.30 pts
  Component 2: H2:H34 formatted to 2 decimal places ('0.00') — 0.10 pts
  Component 3: I2:I34 contain =RANK(Hx,$H$2:$H$34,0) formula — 0.30 pts
  Component 4: J2:J34 contain =IF(Hx>=3.5,"Pass","Fail")     — 0.15 pts
  Component 5: Conditional formatting in J2:J34               — 0.15 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_interview_scorecard_022'


def normalize_formula(f):
    """Normalize formula string for comparison: uppercase, no spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Scorecards' sheet exists
    if 'Scorecards' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Scorecards' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scorecards']

    # Component 1: H2:H34 contain AVERAGE formula (0.30 points)
    # Each row should have =AVERAGE(Cx:Gx) where x is the row number
    try:
        avg_formula_count = 0
        avg_formula_total = 33  # rows 2 through 34

        for row in range(2, 35):
            cell_val = ws.cell(row=row, column=8).value  # column H
            expected_pattern = f'=AVERAGE(C{row}:G{row})'
            norm_val = normalize_formula(str(cell_val) if cell_val else '')
            norm_expected = normalize_formula(expected_pattern)
            if norm_val == norm_expected:
                avg_formula_count += 1

        if avg_formula_count == avg_formula_total:
            print(f"PASS: Component 1 — All {avg_formula_total} rows (H2:H34) contain AVERAGE formula (0.30 pts)")
            total_score += 0.30
        elif avg_formula_count > 0:
            partial = round(0.30 * avg_formula_count / avg_formula_total, 4)
            print(f"PARTIAL: Component 1 — {avg_formula_count}/{avg_formula_total} rows have AVERAGE formula, partial credit: {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No AVERAGE formulas found in H2:H34 (expected =AVERAGE(Cx:Gx))")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: H2:H34 formatted to 2 decimal places (0.10 points)
    try:
        fmt_count = 0
        fmt_total = 33

        for row in range(2, 35):
            cell = ws.cell(row=row, column=8)
            # Accept '0.00' number format
            if cell.number_format in ('0.00', '#,##0.00'):
                fmt_count += 1

        if fmt_count == fmt_total:
            print(f"PASS: Component 2 — All H2:H34 cells have 2-decimal-place number format (0.10 pts)")
            total_score += 0.10
        elif fmt_count > 0:
            partial = round(0.10 * fmt_count / fmt_total, 4)
            print(f"PARTIAL: Component 2 — {fmt_count}/{fmt_total} H-column cells have 2-decimal format, partial credit: {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — H2:H34 cells not formatted to 2 decimal places (found format: {ws.cell(row=2, column=8).number_format!r})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: I2:I34 contain RANK formula (0.30 points)
    # Each row should have =RANK(Hx,$H$2:$H$34,0)
    try:
        rank_formula_count = 0
        rank_formula_total = 33

        for row in range(2, 35):
            cell_val = ws.cell(row=row, column=9).value  # column I
            expected_pattern = f'=RANK(H{row},$H$2:$H$34,0)'
            norm_val = normalize_formula(str(cell_val) if cell_val else '')
            norm_expected = normalize_formula(expected_pattern)
            if norm_val == norm_expected:
                rank_formula_count += 1

        if rank_formula_count == rank_formula_total:
            print(f"PASS: Component 3 — All {rank_formula_total} rows (I2:I34) contain RANK formula (0.30 pts)")
            total_score += 0.30
        elif rank_formula_count > 0:
            partial = round(0.30 * rank_formula_count / rank_formula_total, 4)
            print(f"PARTIAL: Component 3 — {rank_formula_count}/{rank_formula_total} rows have RANK formula, partial credit: {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No RANK formulas found in I2:I34 (expected =RANK(Hx,$H$2:$H$34,0))")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: J2:J34 contain IF formula for Pass/Fail (0.15 points)
    # Each row should have =IF(Hx>=3.5,"Pass","Fail")
    try:
        if_formula_count = 0
        if_formula_total = 33

        for row in range(2, 35):
            cell_val = ws.cell(row=row, column=10).value  # column J
            expected_pattern = f'=IF(H{row}>=3.5,"Pass","Fail")'
            norm_val = normalize_formula(str(cell_val) if cell_val else '')
            norm_expected = normalize_formula(expected_pattern)
            if norm_val == norm_expected:
                if_formula_count += 1

        if if_formula_count == if_formula_total:
            print(f"PASS: Component 4 — All {if_formula_total} rows (J2:J34) contain IF(Pass/Fail) formula (0.15 pts)")
            total_score += 0.15
        elif if_formula_count > 0:
            partial = round(0.15 * if_formula_count / if_formula_total, 4)
            print(f"PARTIAL: Component 4 — {if_formula_count}/{if_formula_total} rows have IF formula, partial credit: {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No IF formulas found in J2:J34 (expected =IF(Hx>=3.5,\"Pass\",\"Fail\"))")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting in J2:J34 (0.15 points)
    # Pass rule: formula J2="Pass", fill green #FF70AD47
    # Fail rule: formula J2="Fail", fill red #FFFF0000, font white #FFFFFFFF
    try:
        pass_rule_count = 0  # counts rules matching the Pass condition (expected: 1)
        fail_rule_count = 0  # counts rules matching the Fail condition (expected: 1)

        for cf_range in ws.conditional_formatting:
            range_str = str(cf_range)
            if 'J2' in range_str and 'J34' in range_str:
                rules = ws.conditional_formatting[cf_range]
                for rule in rules:
                    formula = getattr(rule, 'formula', None)
                    dxf = getattr(rule, 'dxf', None)

                    if formula and dxf:
                        formula_str = normalize_formula(str(formula))

                        # Check for Pass rule: formula references "Pass" and fill is green #70AD47
                        if 'PASS' in formula_str:
                            try:
                                fg_color = dxf.fill.fgColor.rgb
                                if fg_color == 'FF70AD47':
                                    pass_rule_count += 1
                                    print(f"  Found Pass rule: formula={formula}, fill=#{fg_color}")
                            except Exception:
                                pass

                        # Check for Fail rule: formula references "Fail" and fill is red #FF0000
                        if 'FAIL' in formula_str:
                            try:
                                fg_color = dxf.fill.fgColor.rgb
                                font_color = None
                                try:
                                    font_color = dxf.font.color.rgb
                                except Exception:
                                    pass
                                if fg_color == 'FFFF0000':
                                    fail_rule_count += 1
                                    print(f"  Found Fail rule: formula={formula}, fill=#{fg_color}, font=#{font_color}")
                            except Exception:
                                pass

        if pass_rule_count >= 1 and fail_rule_count >= 1:
            print(f"PASS: Component 5 — Conditional formatting for Pass (green) and Fail (red) found in J2:J34 (0.15 pts)")
            total_score += 0.15
        elif pass_rule_count >= 1 or fail_rule_count >= 1:
            print(f"PARTIAL: Component 5 — Only one conditional formatting rule matched (Pass={pass_rule_count}, Fail={fail_rule_count}), partial credit: 0.075 pts")
            total_score += 0.075
        else:
            print(f"FAIL: Component 5 — No matching conditional formatting rules found in J2:J34")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
