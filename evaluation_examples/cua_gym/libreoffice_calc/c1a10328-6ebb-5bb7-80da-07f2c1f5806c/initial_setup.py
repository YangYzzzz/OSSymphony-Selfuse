"""
Initial Setup: HR Interview Scorecard - candidate scores without averages/ranks/recommend
Task ID: calc_hr_interview_scorecard_022
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_interview_scorecard_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Scorecards'

    # --- Headers ---
    headers = [
        'Candidate', 'Position', 'Technical', 'Communication',
        'Problem Solving', 'Culture Fit', 'Leadership',
        'Average Score', 'Rank', 'Recommend'
    ]
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 24
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 16
    for col_letter in ['H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 16

    # --- Candidate data (33 rows, realistic names, positions, scores 1-5) ---
    candidates = [
        # Name, Position, Technical, Communication, Problem Solving, Culture Fit, Leadership
        ('Aiden Park',          'Software Engineer',      5, 4, 5, 4, 3),
        ('Brianna Torres',      'Marketing Manager',      3, 5, 4, 5, 4),
        ('Carlos Reyes',        'Data Analyst',           4, 3, 5, 3, 2),
        ('Diana Nguyen',        'UX Designer',            3, 4, 4, 5, 3),
        ('Ethan Kowalski',      'Product Manager',        4, 5, 4, 4, 5),
        ('Fatima Al-Hassan',    'Financial Analyst',      5, 3, 4, 3, 3),
        ('George Mitchell',     'Operations Lead',        3, 4, 3, 4, 4),
        ('Hannah Osei',         'HR Specialist',          2, 5, 3, 5, 4),
        ('Ivan Petrov',         'Software Engineer',      5, 3, 5, 3, 4),
        ('Julia Schneider',     'Project Manager',        4, 4, 4, 4, 5),
        ('Kevin O\'Brien',      'Data Analyst',           3, 2, 4, 3, 2),
        ('Layla Ahmed',         'Marketing Coordinator',  4, 5, 3, 5, 3),
        ('Marcus Thompson',     'DevOps Engineer',        5, 3, 5, 4, 3),
        ('Nadia Kostadinova',   'Business Analyst',       4, 4, 4, 3, 4),
        ('Oscar Fernandez',     'Product Manager',        3, 4, 4, 4, 4),
        ('Priya Sharma',        'Data Scientist',         5, 4, 5, 4, 4),
        ('Quinn Blackwell',     'UX Researcher',          3, 5, 3, 5, 3),
        ('Rachel Kim',          'Financial Analyst',      4, 4, 4, 4, 3),
        ('Samuel Greene',       'Software Engineer',      5, 3, 4, 3, 5),
        ('Tanya Wilson',        'HR Manager',             3, 5, 3, 5, 5),
        ('Umar Hussain',        'Cloud Architect',        5, 3, 5, 3, 4),
        ('Vera Johansson',      'Operations Analyst',     4, 4, 3, 4, 3),
        ('William Chen',        'Marketing Manager',      3, 5, 4, 5, 4),
        ('Xiomara Diaz',        'Data Analyst',           4, 3, 5, 3, 3),
        ('Yasmine Belkacem',    'Project Coordinator',    3, 4, 3, 4, 2),
        ('Zachary Holmes',      'Software Engineer',      4, 4, 5, 4, 4),
        ('Amara Okonkwo',       'UX Designer',            3, 5, 4, 5, 4),
        ('Ben Nakamura',        'DevOps Engineer',        5, 3, 5, 3, 3),
        ('Clara Monteiro',      'Business Analyst',       4, 4, 4, 4, 4),
        ('Daniel Hofmann',      'Cloud Architect',        5, 4, 5, 4, 5),
        ('Elena Vasquez',       'HR Specialist',          2, 5, 3, 5, 3),
        ('Frank Oduya',         'Operations Lead',        4, 4, 4, 4, 4),
        ('Grace Liu',           'Data Scientist',         5, 4, 5, 4, 5),
    ]

    # Score fill colors for alternating rows
    fill_even = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    fill_odd  = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')

    for r, (name, pos, tech, comm, prob, cult, lead) in enumerate(candidates, 2):
        row_fill = fill_even if r % 2 == 0 else fill_odd
        vals = [name, pos, tech, comm, prob, cult, lead]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal='center' if c >= 3 else 'left',
                                       vertical='center')
        # Columns H, I, J intentionally left empty

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, 35):
        ws.row_dimensions[r].height = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
