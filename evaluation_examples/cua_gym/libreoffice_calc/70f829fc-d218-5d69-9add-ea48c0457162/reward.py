"""
Reward Script: Format large numbers in scientific notation using '0.00E+00'
Task ID: calc_lf_064
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): B2 has number_format '0.00E+00'
  Component 2 (0.35): B3 has number_format '0.00E+00'
  Component 3 (0.30): B4 has number_format '0.00E+00'
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_064'

EXPECTED_FORMAT = '0.00E+00'


def persist_app_state(domain: str):
    """Try to save any unsaved LibreOffice edits."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Science' sheet must exist
    if 'Science' not in wb.sheetnames:
        print(f"FAIL: 'Science' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Science']

    # Component 1: B2 number format is '0.00E+00' (0.35 points)
    try:
        fmt_b2 = ws['B2'].number_format
        if fmt_b2 == EXPECTED_FORMAT:
            print(f"PASS: Component 1 — B2 number_format is '{fmt_b2}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — B2 number_format expected '{EXPECTED_FORMAT}', found '{fmt_b2}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: B3 number format is '0.00E+00' (0.35 points)
    try:
        fmt_b3 = ws['B3'].number_format
        if fmt_b3 == EXPECTED_FORMAT:
            print(f"PASS: Component 2 — B3 number_format is '{fmt_b3}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — B3 number_format expected '{EXPECTED_FORMAT}', found '{fmt_b3}'")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: B4 number format is '0.00E+00' (0.30 points)
    try:
        fmt_b4 = ws['B4'].number_format
        if fmt_b4 == EXPECTED_FORMAT:
            print(f"PASS: Component 3 — B4 number_format is '{fmt_b4}' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — B4 number_format expected '{EXPECTED_FORMAT}', found '{fmt_b4}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
