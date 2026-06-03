"""
Initial Setup: Format large numbers in scientific notation
Task ID: calc_lf_064
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Science ---
    ws = wb.active
    ws.title = 'Science'

    # Headers
    ws.cell(row=1, column=1, value='Measurement')
    ws.cell(row=1, column=2, value='Value')

    # Data rows with large scientific values
    data = [
        ['Distance', 149597870700],
        ['Mass', 5972000000],
        ['Speed', 299792458],
    ]
    for r, (label, val) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=label)
        ws.cell(row=r, column=2, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
