"""
Reward Script: Product mix analysis — percentage of total revenue
Task ID: calc_sales_042
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35) — B6 contains a SUM formula over B2:B5
  Component 2 (0.40) — C2:C5 contain percentage formulas (=Bx/B$6)
  Component 3 (0.25) — C2:C5 are formatted as percentage
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_042'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'ProductMix' sheet must exist
    if 'ProductMix' not in wb.sheetnames:
        print("FAIL: Sheet 'ProductMix' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ProductMix']

    # ---------------------------------------------------------------
    # Component 1: B6 contains a SUM formula over B2:B5 (0.35 points)
    # Initial: B6 is None -> FAIL
    # Golden:  B6 = "=SUM(B2:B5)" -> PASS
    # ---------------------------------------------------------------
    try:
        b6_val = ws['B6'].value
        if isinstance(b6_val, str):
            # Normalize: remove spaces, uppercase
            normalized = b6_val.upper().replace(" ", "")
            # Accept =SUM(B2:B5) or equivalent ranges
            if re.match(r'^=SUM\(B2:B5\)$', normalized):
                print(f"PASS: Component 1 — B6 has SUM formula: {b6_val} (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 1 — B6 has formula but not =SUM(B2:B5): {b6_val}")
        else:
            # Also accept if the value is the correct number (1000000) when saved by Calc
            if b6_val is not None and abs(float(b6_val) - 1000000) < 0.01:
                # Check if there was a formula that got evaluated
                # Load with data_only=False already done above, so if value is numeric,
                # the formula was replaced. Still accept computed result.
                print(f"PASS: Component 1 — B6 has computed SUM value: {b6_val} (0.35 pts)")
                total_score += 0.35
            else:
                print(f"FAIL: Component 1 — B6 expected SUM formula or 1000000, found: {b6_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: C2:C5 contain percentage formulas referencing B/B6
    # Initial: C2:C5 are all None -> FAIL
    # Golden:  C2=B2/B$6, C3=B3/B$6, C4=B4/B$6, C5=B5/B$6 -> PASS
    # We award 0.10 per correct cell (4 cells = 0.40 total)
    # ---------------------------------------------------------------
    try:
        expected_formulas = {
            'C2': ('B2', 'B$6'),
            'C3': ('B3', 'B$6'),
            'C4': ('B4', 'B$6'),
            'C5': ('B5', 'B$6'),
        }
        # Also accept computed percentage values as valid
        expected_values = {
            'C2': 0.45,
            'C3': 0.28,
            'C4': 0.175,
            'C5': 0.095,
        }

        comp2_score = 0.0
        for cell_ref, (numerator, denominator) in expected_formulas.items():
            cell_val = ws[cell_ref].value
            if isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                # Accept =B2/B$6 or =B2/$B$6 patterns
                # Escape $ in denominator for regex
                denom_escaped = denominator.upper().replace('$', '\\$')
                pattern = f'^={numerator.upper()}/\\$?{denom_escaped}$'
                if re.match(pattern, normalized):
                    print(f"PASS: {cell_ref} has correct formula: {cell_val}")
                    comp2_score += 0.10
                else:
                    print(f"FAIL: {cell_ref} expected ={numerator}/{denominator}, found: {cell_val}")
            elif cell_val is not None:
                # Accept computed decimal value
                try:
                    exp_val = expected_values[cell_ref]
                    if abs(float(cell_val) - exp_val) < 0.005:
                        print(f"PASS: {cell_ref} has correct computed percentage: {cell_val}")
                        comp2_score += 0.10
                    else:
                        print(f"FAIL: {cell_ref} expected ~{exp_val}, found: {cell_val}")
                except (ValueError, TypeError):
                    print(f"FAIL: {cell_ref} unexpected value type: {cell_val}")
            else:
                print(f"FAIL: {cell_ref} is empty (None)")

        if comp2_score > 0:
            print(f"Component 2 subtotal: {comp2_score}/0.40")
            total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: C2:C5 formatted as percentage (0.25 points)
    # Initial: C2:C5 have 'General' format -> FAIL
    # Golden:  C2:C5 have '0.0%' or similar percentage format -> PASS
    # Award 0.0625 per cell formatted as percentage
    # ---------------------------------------------------------------
    try:
        comp3_score = 0.0
        for cell_ref in ['C2', 'C3', 'C4', 'C5']:
            fmt = ws[cell_ref].number_format
            # Check if format contains '%' (covers 0%, 0.0%, 0.00%, etc.)
            if fmt and '%' in fmt:
                print(f"PASS: {cell_ref} formatted as percentage: {fmt}")
                comp3_score += 0.0625
            else:
                print(f"FAIL: {cell_ref} not percentage format: {fmt}")

        if comp3_score > 0:
            print(f"Component 3 subtotal: {comp3_score:.4f}/0.25")
            total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
