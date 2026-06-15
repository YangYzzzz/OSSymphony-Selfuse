"""
Initial Setup: Product mix analysis spreadsheet with revenue data
Task ID: calc_sales_042
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProductMix ---
    ws = wb.active
    ws.title = 'ProductMix'

    # Headers
    headers = ['Product', 'Revenue', '% of Total']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data rows — revenue figures, NO formulas, NO percentages
    data = [
        ['Product A', 450000],
        ['Product B', 280000],
        ['Product C', 175000],
        ['Product D', 95000],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0]).border = thin_border
        rev_cell = ws.cell(row=r, column=2, value=row_data[1])
        rev_cell.number_format = '$#,##0'
        rev_cell.border = thin_border
        # Column C left empty — agent must fill in percentage formulas
        ws.cell(row=r, column=3).border = thin_border

    # Total label row — B6 left empty (agent must add SUM formula)
    total_cell = ws.cell(row=6, column=1, value='Total')
    total_cell.font = Font(name='Calibri', size=11, bold=True)
    total_cell.border = thin_border
    ws.cell(row=6, column=2).border = thin_border
    ws.cell(row=6, column=2).number_format = '$#,##0'
    ws.cell(row=6, column=3).border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
