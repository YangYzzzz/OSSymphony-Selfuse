"""
Reward Script: Write a shell script to resize images with ImageMagick, log to CSV, open in LibreOffice Calc with AVERAGE formula
Task ID: osworld_multi_apps_code_batch_terminal_010
Domain: libreoffice_calc (multi-app: terminal + LibreOffice Calc)
Scoring:
  Component 1: Shell script exists at /home/user/scripts/media_processor.sh (0.20 pts)
  Component 2: Resized directory exists with 15 resized images (0.20 pts)
  Component 3: CSV file exists with header + 15 data rows and correct columns (0.25 pts)
  Component 4: XLSX file exists with 'Resize Log' sheet and 15 data rows (0.20 pts)
  Component 5: XLSX has AVERAGE formula cell computing size reduction percentage (0.15 pts)
"""

import os
import csv

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_code_batch_terminal_010'

SCRIPT_PATH = '/home/user/scripts/media_processor.sh'
RESIZED_DIR = '/home/user/media/resized'
CSV_PATH = '/home/user/media/resize_log.csv'
XLSX_PATH = f'/home/user/{TASK_ID}.xlsx'

EXPECTED_IMAGE_COUNT = 15
EXPECTED_CSV_ROWS = 15  # data rows (excluding header)


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Component 1: Shell script exists at /home/user/scripts/media_processor.sh (0.20 points)
    # This FAILS on initial (scripts/ is empty) and PASSES on golden
    try:
        script_exists = os.path.isfile(SCRIPT_PATH)
        if script_exists:
            # Check it is non-empty and contains expected keywords for image resizing
            with open(SCRIPT_PATH, 'r') as f:
                script_content = f.read()
            has_convert = 'convert' in script_content or 'ImageMagick' in script_content.lower() or 'resize' in script_content.lower()
            has_csv_ref = 'csv' in script_content.lower() or 'resize_log' in script_content
            if has_convert and has_csv_ref:
                print(f"PASS: Component 1 — shell script exists with image resize + CSV content (0.20 pts)")
                total_score += 0.20
            elif script_exists:
                print(f"PASS (partial): Component 1 — shell script exists but missing expected content (0.10 pts)")
                total_score += 0.10
        else:
            print(f"FAIL: Component 1 — shell script not found at {SCRIPT_PATH}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Resized directory exists with 15 resized image files (0.20 points)
    # This FAILS on initial (resized/ does not exist) and PASSES on golden
    try:
        if not os.path.isdir(RESIZED_DIR):
            print(f"FAIL: Component 2 — resized directory does not exist: {RESIZED_DIR}")
        else:
            resized_files = [f for f in os.listdir(RESIZED_DIR) if f.lower().endswith(('.jpg', '.png', '.jpeg'))]
            count = len(resized_files)
            if count == EXPECTED_IMAGE_COUNT:
                print(f"PASS: Component 2 — resized directory exists with exactly {count} image files (0.20 pts)")
                total_score += 0.20
            elif count > 0:
                partial = round(0.20 * count / EXPECTED_IMAGE_COUNT, 3)
                print(f"PARTIAL: Component 2 — resized directory has {count}/{EXPECTED_IMAGE_COUNT} images ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — resized directory exists but has no image files")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: CSV file exists with header + 15 data rows and correct columns (0.25 points)
    # This FAILS on initial (no CSV file) and PASSES on golden
    try:
        if not os.path.isfile(CSV_PATH):
            print(f"FAIL: Component 3 — CSV file not found at {CSV_PATH}")
        else:
            with open(CSV_PATH, 'r', newline='') as csvf:
                reader = csv.DictReader(csvf)
                rows = list(reader)
                fieldnames = reader.fieldnames or []

            # Check required columns
            required_cols = {'filename', 'original_size', 'new_size'}
            # Case-insensitive check for column names
            actual_cols_lower = {c.lower().replace(' ', '_') for c in fieldnames}
            cols_ok = required_cols.issubset(actual_cols_lower)

            row_count = len(rows)
            if cols_ok and row_count == EXPECTED_CSV_ROWS:
                # Verify data rows have numeric size values
                valid_rows = 0
                for row in rows:
                    try:
                        orig = int(list(row.values())[1])
                        new = int(list(row.values())[2])
                        if orig > 0 and new > 0:
                            valid_rows += 1
                    except (ValueError, IndexError):
                        pass
                if valid_rows == EXPECTED_CSV_ROWS:
                    print(f"PASS: Component 3 — CSV has header + {row_count} data rows with required columns and numeric values (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"PARTIAL: Component 3 — CSV has correct structure but only {valid_rows}/{EXPECTED_CSV_ROWS} rows have valid numeric sizes (0.15 pts)")
                    total_score += 0.15
            elif cols_ok and row_count > 0:
                partial = round(0.25 * row_count / EXPECTED_CSV_ROWS, 3)
                print(f"PARTIAL: Component 3 — CSV has correct columns but {row_count}/{EXPECTED_CSV_ROWS} data rows ({partial} pts)")
                total_score += partial
            elif row_count == EXPECTED_CSV_ROWS:
                print(f"PARTIAL: Component 3 — CSV has {row_count} rows but missing required columns: {required_cols - actual_cols_lower} (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 — CSV has {row_count} data rows, expected {EXPECTED_CSV_ROWS}, columns: {fieldnames}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: XLSX file exists with 'Resize Log' sheet and 15 data rows (0.20 points)
    # This FAILS on initial (no XLSX) and PASSES on golden
    try:
        import openpyxl
        if not os.path.isfile(XLSX_PATH):
            print(f"FAIL: Component 4 — XLSX file not found at {XLSX_PATH}")
        else:
            wb = openpyxl.load_workbook(XLSX_PATH)
            # Accept either exact sheet name or first sheet that contains data
            sheet_names_lower = [s.lower() for s in wb.sheetnames]
            has_resize_log_sheet = any('resize' in s for s in sheet_names_lower)

            if not has_resize_log_sheet and len(wb.sheetnames) == 0:
                print(f"FAIL: Component 4 — XLSX has no sheets (sheets: {wb.sheetnames})")
            else:
                # Use the resize-related sheet, or first sheet
                ws = None
                for sname in wb.sheetnames:
                    if 'resize' in sname.lower():
                        ws = wb[sname]
                        break
                if ws is None:
                    ws = wb.active

                # Count data rows (rows with actual filename content)
                data_rows = 0
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0] is not None and str(row[0]).strip() != '':
                        # Check it looks like a filename (has . extension)
                        if '.' in str(row[0]):
                            data_rows += 1

                if data_rows == EXPECTED_CSV_ROWS:
                    print(f"PASS: Component 4 — XLSX has sheet '{ws.title}' with {data_rows} data rows (0.20 pts)")
                    total_score += 0.20
                elif data_rows > 0:
                    partial = round(0.20 * data_rows / EXPECTED_CSV_ROWS, 3)
                    print(f"PARTIAL: Component 4 — XLSX has {data_rows}/{EXPECTED_CSV_ROWS} data rows ({partial} pts)")
                    total_score += partial
                else:
                    print(f"FAIL: Component 4 — XLSX sheet '{ws.title}' has no data rows")
    except ImportError:
        print("ERROR: Component 4 — openpyxl not available")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: XLSX has AVERAGE formula cell computing size reduction percentage (0.15 points)
    # This FAILS on initial (no XLSX) and PASSES on golden
    # Expected: a cell containing =AVERAGE(...) formula referencing reduction % column
    try:
        import openpyxl
        if not os.path.isfile(XLSX_PATH):
            print(f"FAIL: Component 5 — XLSX file not found at {XLSX_PATH}")
        else:
            wb_formulas = openpyxl.load_workbook(XLSX_PATH)
            ws_f = None
            for sname in wb_formulas.sheetnames:
                if 'resize' in sname.lower():
                    ws_f = wb_formulas[sname]
                    break
            if ws_f is None:
                ws_f = wb_formulas.active

            # Search all cells for AVERAGE formula
            found_average = False
            average_cell_ref = None
            avg_references_column = False

            for row in ws_f.iter_rows():
                for cell in row:
                    val = cell.value
                    if isinstance(val, str) and val.strip().upper().startswith('=AVERAGE('):
                        found_average = True
                        average_cell_ref = cell.coordinate
                        # Check that formula references a column range (e.g., D2:D16)
                        # This verifies it's computing average over the data rows
                        import re
                        range_match = re.search(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', val.upper())
                        if range_match:
                            avg_references_column = True
                        break
                if found_average:
                    break

            if found_average and avg_references_column:
                print(f"PASS: Component 5 — XLSX has AVERAGE formula at {average_cell_ref} referencing column range (0.15 pts)")
                total_score += 0.15
            elif found_average:
                print(f"PASS (partial): Component 5 — XLSX has AVERAGE formula at {average_cell_ref} (0.10 pts)")
                total_score += 0.10
            else:
                # Also search for a label indicating average was placed nearby
                found_avg_label = False
                for row in ws_f.iter_rows(values_only=True):
                    for cell_val in row:
                        if cell_val and 'average' in str(cell_val).lower():
                            found_avg_label = True
                            break
                    if found_avg_label:
                        break
                if found_avg_label:
                    print(f"FAIL: Component 5 — XLSX has 'average' label but no AVERAGE formula found")
                else:
                    print(f"FAIL: Component 5 — No AVERAGE formula found in XLSX")
    except ImportError:
        print("ERROR: Component 5 — openpyxl not available")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


verify_task()
