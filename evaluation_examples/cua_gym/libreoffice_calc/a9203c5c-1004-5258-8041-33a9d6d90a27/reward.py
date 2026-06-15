"""
Reward Script: Custom ordinal number format on B2:B6
Task ID: calc_lf_095
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Custom format applied to ALL cells B2:B6
  Component 2 (0.3): Format contains correct conditional sections for 1 and 2
  Component 3 (0.3): Underlying numeric values preserved in B2:B6
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_095'

# The expected custom number format string
EXPECTED_FORMAT = '[=1]0"st";[=2]0"nd";0"th"'

# The cells that should have the custom format applied
TARGET_CELLS = ['B2', 'B3', 'B4', 'B5', 'B6']

# Expected underlying numeric values (must be preserved)
EXPECTED_VALUES = {
    'B2': 1,
    'B3': 2,
    'B4': 3,
    'B5': 10,
    'B6': 21,
}


def normalize_format(fmt):
    """Normalize a number format string for comparison.
    Strips whitespace and lowercases for flexible matching.
    """
    if fmt is None:
        return ''
    return fmt.replace(' ', '').lower()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet 'Rankings' exists
    if 'Rankings' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Rankings' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Rankings']

    # Component 1: Custom format applied to ALL cells B2:B6 (0.4 points)
    # This is the primary task change - format must be non-General on all target cells
    try:
        formatted_count = 0
        for cell_ref in TARGET_CELLS:
            cell = ws[cell_ref]
            nf = cell.number_format
            if nf is not None and nf != 'General':
                formatted_count += 1
                print(f"  {cell_ref}: has custom format '{nf}'")
            else:
                print(f"  {cell_ref}: still has default format '{nf}'")

        if formatted_count == len(TARGET_CELLS):
            print(f"PASS: Component 1 - All {len(TARGET_CELLS)} cells have custom format (0.4 pts)")
            total_score += 0.4
        elif formatted_count > 0:
            partial = round(0.4 * formatted_count / len(TARGET_CELLS), 2)
            print(f"PARTIAL: Component 1 - {formatted_count}/{len(TARGET_CELLS)} cells formatted ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No cells have custom format applied")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Format string matches the expected ordinal pattern (0.3 points)
    # Must contain conditional sections for =1 and =2 with ordinal suffixes
    try:
        correct_format_count = 0
        expected_norm = normalize_format(EXPECTED_FORMAT)

        for cell_ref in TARGET_CELLS:
            cell = ws[cell_ref]
            actual_norm = normalize_format(cell.number_format)
            if actual_norm == expected_norm:
                correct_format_count += 1
            else:
                print(f"  {cell_ref}: format mismatch - expected '{EXPECTED_FORMAT}', got '{cell.number_format}'")

        if correct_format_count == len(TARGET_CELLS):
            print(f"PASS: Component 2 - All cells have exact ordinal format (0.3 pts)")
            total_score += 0.3
        elif correct_format_count > 0:
            partial = round(0.3 * correct_format_count / len(TARGET_CELLS), 2)
            print(f"PARTIAL: Component 2 - {correct_format_count}/{len(TARGET_CELLS)} cells have correct format ({partial} pts)")
            total_score += partial
        else:
            # Secondary check: does the format at least contain ordinal-related patterns?
            ordinal_cell_count = sum(
                1 for c in TARGET_CELLS
                if '[=1]' in (ws[c].number_format or '') and '[=2]' in (ws[c].number_format or '')
                   and 'st' in (ws[c].number_format or '').lower() and 'nd' in (ws[c].number_format or '').lower()
            )
            if ordinal_cell_count > 0:
                print(f"PARTIAL: Component 2 - Format has ordinal sections but not exact match (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 2 - Format does not match expected ordinal pattern")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Underlying numeric values preserved in B2:B6 (0.3 points)
    # Values must still be numeric (not converted to text like "1st")
    # AND the format must be non-General (combines preservation check with task-change anchor)
    try:
        preserved_count = 0
        for cell_ref, expected_val in EXPECTED_VALUES.items():
            cell = ws[cell_ref]
            val = cell.value
            nf = cell.number_format
            # Value must match AND format must be non-General (task-change anchor)
            if isinstance(val, (int, float)) and val == expected_val and nf != 'General':
                preserved_count += 1
            else:
                print(f"  {cell_ref}: value={val!r} (expected {expected_val}), format='{nf}'")

        if preserved_count == len(EXPECTED_VALUES):
            print(f"PASS: Component 3 - All values preserved as numbers with custom format (0.3 pts)")
            total_score += 0.3
        elif preserved_count > 0:
            partial = round(0.3 * preserved_count / len(EXPECTED_VALUES), 2)
            print(f"PARTIAL: Component 3 - {preserved_count}/{len(EXPECTED_VALUES)} values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - Values not preserved or format not applied")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
