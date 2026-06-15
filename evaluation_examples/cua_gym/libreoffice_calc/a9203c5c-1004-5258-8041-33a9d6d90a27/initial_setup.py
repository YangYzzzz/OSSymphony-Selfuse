"""
Initial Setup: Create a spreadsheet with numeric ranking data for custom ordinal formatting task.
Task ID: calc_lf_095
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Rankings ---
    ws = wb.active
    ws.title = 'Rankings'

    # Headers
    ws['A1'] = 'Position'
    ws['B1'] = 'Display'

    # Style headers
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in [ws['A1'], ws['B1']]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Data rows - numeric values only (NO custom format applied - that's the task)
    rankings_data = [
        [1, 1],
        [2, 2],
        [3, 3],
        [4, 10],
        [5, 21],
    ]
    for r, (pos, val) in enumerate(rankings_data, 2):
        ws.cell(row=r, column=1, value=pos)
        ws.cell(row=r, column=2, value=val)
        # Center-align data cells
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')

    # Add a second sheet with context data to make the workbook realistic
    ws2 = wb.create_sheet('Competition Info')
    ws2['A1'] = 'Competitor'
    ws2['B1'] = 'Score'
    ws2['C1'] = 'Region'

    competitors = [
        ['Elena Vasquez', 94.5, 'North America'],
        ['Takeshi Yamamoto', 91.2, 'Asia Pacific'],
        ['Friedrich Weber', 88.7, 'Europe'],
        ['Priya Sharma', 85.3, 'Asia Pacific'],
        ['Liam O\'Brien', 82.1, 'Europe'],
        ['Amara Okafor', 79.8, 'Africa'],
        ['Sophie Laurent', 77.4, 'Europe'],
        ['Carlos Mendez', 74.9, 'South America'],
        ['Jin-soo Park', 72.6, 'Asia Pacific'],
        ['Hannah Mitchell', 70.1, 'North America'],
    ]
    for cell in [ws2['A1'], ws2['B1'], ws2['C1']]:
        cell.font = Font(bold=True)

    for r, row_data in enumerate(competitors, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
