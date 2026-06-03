"""
Initial Setup: Semester Course Planner
Task ID: calc_edu_semester_planner_056
Domain: libreoffice_calc

Creates a semester course planner spreadsheet with course codes, names, credits,
and type (Core Requirement, Elective, Free Elective). Cost column is empty.
Row 10 has summary labels but no formulas. No conditional formatting applied.
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_semester_planner_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CoursePlan'

    # --- Headers (Row 1) ---
    headers = ['Course Code', 'Course Name', 'Credits', 'Type', 'Cost']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # --- Course data (Rows 2-8) ---
    # 7 planned courses mixing Core Requirement, Elective, and Free Elective
    courses = [
        ('CS101',  'Introduction to Computer Science',  3, 'Core Requirement'),
        ('MATH201', 'Calculus II',                       4, 'Core Requirement'),
        ('ENG110',  'Academic Writing',                  3, 'Core Requirement'),
        ('HIST150',  'World History Survey',              3, 'Elective'),
        ('PHYS101', 'Physics for Engineers',             4, 'Core Requirement'),
        ('ART200',  'Introduction to Digital Media',     2, 'Elective'),
        ('PE105',   'Health and Wellness',               1, 'Free Elective'),
    ]

    for r, (code, name, credits, ctype) in enumerate(courses, 2):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=credits)
        ws.cell(row=r, column=4, value=ctype)
        # Column E (Cost) intentionally left EMPTY — to be filled by agent

    # --- Row 10: Summary labels (no formulas) ---
    ws['A10'] = 'Total Credits'
    ws['B10'] = 'Total Cost'
    ws['C10'] = 'Full Time Status'
    # B10, C10, D10 intentionally left blank (no formulas)

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
