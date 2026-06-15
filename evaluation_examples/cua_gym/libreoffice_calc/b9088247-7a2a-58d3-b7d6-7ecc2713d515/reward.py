"""
Reward Script: Build a semester course planner with credit hour calculations, tuition cost,
full-time status verification, and color-coded conditional formatting.
Task ID: calc_edu_semester_planner_056
Domain: libreoffice_calc
Scoring:
  - Component 1: Cost formulas in E2:E8 (=Cx*350) with currency format ($#,##0.00) — 0.30 pts
  - Component 2: Summary formulas in B10 (=SUM), C10 (=B10*350), D10 (=IF full-time) — 0.40 pts
  - Component 3: Row 10 is bold — 0.10 pts
  - Component 4: Conditional formatting (3 rules: Core Req blue, Elective green, Free Elective yellow) — 0.20 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_semester_planner_056'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet exists
    if 'CoursePlan' not in wb.sheetnames:
        print("CRITICAL: Sheet 'CoursePlan' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CoursePlan']

    # Component 1: Cost formulas in E2:E8 with currency format (0.30 points)
    # Each of the 7 course rows (2-8) should have =Cx*350 in column E, formatted as currency
    try:
        formula_count = 0
        currency_count = 0
        for row in range(2, 9):
            cell = ws.cell(row=row, column=5)  # Column E
            expected_formula = f'=C{row}*350'
            # Check formula (case-insensitive, normalize spaces)
            if cell.value and isinstance(cell.value, str):
                normalized = cell.value.upper().replace(' ', '')
                expected_norm = expected_formula.upper().replace(' ', '')
                if normalized == expected_norm:
                    formula_count += 1
            # Check currency number format
            if cell.number_format and '$' in cell.number_format:
                currency_count += 1

        if formula_count == 7 and currency_count == 7:
            print(f"PASS: Component 1 — All 7 cost formulas (=Cx*350) present with currency format (0.30 pts)")
            total_score += 0.30
        elif formula_count >= 4:
            # Partial credit: at least 4 correct formulas
            partial = round(0.15 * (formula_count / 7), 2)
            print(f"PARTIAL: Component 1 — {formula_count}/7 cost formulas correct, {currency_count}/7 with currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/7 cost formulas in E2:E8. Expected =C2*350 through =C8*350")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Summary formulas in row 10 (0.40 points)
    # B10: =SUM(C2:C8), C10: =B10*350 with currency, D10: =IF(B10>=12,"Full-Time","Part-Time")
    try:
        b10 = ws.cell(row=10, column=2)
        c10 = ws.cell(row=10, column=3)
        d10 = ws.cell(row=10, column=4)

        # Check B10 is SUM formula
        b10_ok = False
        if b10.value and isinstance(b10.value, str):
            b10_norm = b10.value.upper().replace(' ', '')
            if b10_norm == '=SUM(C2:C8)':
                b10_ok = True

        # Check C10 is =B10*350 with currency format
        c10_ok = False
        if c10.value and isinstance(c10.value, str):
            c10_norm = c10.value.upper().replace(' ', '')
            if c10_norm == '=B10*350' and c10.number_format and '$' in c10.number_format:
                c10_ok = True

        # Check D10 is IF formula for full-time status
        d10_ok = False
        if d10.value and isinstance(d10.value, str):
            d10_norm = d10.value.upper().replace(' ', '')
            # Accept variations of the IF formula
            if 'IF(B10>=12' in d10_norm and 'FULL-TIME' in d10_norm:
                d10_ok = True

        component2_score = 0.0
        if b10_ok:
            component2_score += 0.15
            print(f"PASS: Component 2a — B10 has =SUM(C2:C8) (0.15 pts)")
        else:
            print(f"FAIL: Component 2a — B10 expected =SUM(C2:C8), found: {repr(b10.value)}")

        if c10_ok:
            component2_score += 0.15
            print(f"PASS: Component 2b — C10 has =B10*350 with currency format (0.15 pts)")
        else:
            print(f"FAIL: Component 2b — C10 expected =B10*350 with currency, found: value={repr(c10.value)}, fmt={c10.number_format}")

        if d10_ok:
            component2_score += 0.10
            print(f"PASS: Component 2c — D10 has IF full-time status formula (0.10 pts)")
        else:
            print(f"FAIL: Component 2c — D10 expected IF(B10>=12,\"Full-Time\",\"Part-Time\"), found: {repr(d10.value)}")

        total_score += component2_score

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row 10 is bold (0.10 points)
    try:
        bold_count = 0
        # Check cells A10-D10 for bold (cells that have values)
        cells_to_check = [(10, 1), (10, 2), (10, 3), (10, 4)]  # A10, B10, C10, D10
        for r, c in cells_to_check:
            cell = ws.cell(row=r, column=c)
            if cell.font and cell.font.bold:
                bold_count += 1

        if bold_count >= 3:
            print(f"PASS: Component 3 — Row 10 is bold ({bold_count}/4 key cells bold) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 3 — Only {bold_count}/4 key cells in row 10 are bold")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting with 3 rules for course types (0.20 points)
    # Rule 1: Core Requirement -> blue (#4472C4 = FF4472C4 in ARGB)
    # Rule 2: Elective -> green (#70AD47 = FF70AD47 in ARGB)
    # Rule 3: Free Elective -> yellow (#FFFF00 = FFFFFF00 in ARGB)
    try:
        cf_rules = ws.conditional_formatting
        cf_list = list(cf_rules)

        # Collect all rules across all CF ranges
        all_rules = []
        for cf_range in cf_list:
            for rule in cf_range.rules:
                all_rules.append(rule)

        # Check for the 3 required CF rules
        found_core = False
        found_elective = False
        found_free_elective = False

        for rule in all_rules:
            if rule.type == 'expression' and rule.formula:
                formula_str = str(rule.formula).upper()
                # Check color if dxf is available
                dxf_color = None
                try:
                    if rule.dxf and rule.dxf.fill and rule.dxf.fill.fgColor:
                        dxf_color = rule.dxf.fill.fgColor.rgb
                except Exception:
                    pass

                # Core Requirement check (blue #4472C4)
                if 'CORE REQUIREMENT' in formula_str:
                    if dxf_color and '4472C4' in dxf_color.upper():
                        found_core = True
                        print(f"PASS: CF rule for 'Core Requirement' with blue color found")
                    elif dxf_color is None:
                        # Formula exists but can't verify color
                        found_core = True
                        print(f"PASS: CF rule for 'Core Requirement' found (color check skipped)")
                    else:
                        print(f"FAIL: CF rule for 'Core Requirement' has wrong color: {dxf_color}")

                # Elective check (green #70AD47) - must not match Free Elective
                elif 'FREE ELECTIVE' in formula_str:
                    if dxf_color and 'FFFF00' in dxf_color.upper():
                        found_free_elective = True
                        print(f"PASS: CF rule for 'Free Elective' with yellow color found")
                    elif dxf_color is None:
                        found_free_elective = True
                        print(f"PASS: CF rule for 'Free Elective' found (color check skipped)")
                    else:
                        print(f"FAIL: CF rule for 'Free Elective' has wrong color: {dxf_color}")

                elif 'ELECTIVE' in formula_str and 'FREE' not in formula_str:
                    if dxf_color and '70AD47' in dxf_color.upper():
                        found_elective = True
                        print(f"PASS: CF rule for 'Elective' with green color found")
                    elif dxf_color is None:
                        found_elective = True
                        print(f"PASS: CF rule for 'Elective' found (color check skipped)")
                    else:
                        print(f"FAIL: CF rule for 'Elective' has wrong color: {dxf_color}")

        cf_score = 0.0
        if found_core:
            cf_score += 0.07
        if found_elective:
            cf_score += 0.07
        if found_free_elective:
            cf_score += 0.06

        if found_core and found_elective and found_free_elective:
            print(f"PASS: Component 4 — All 3 conditional formatting rules present (0.20 pts)")
            total_score += 0.20
        elif cf_score > 0:
            print(f"PARTIAL: Component 4 — {int(cf_score/0.07 + 0.5)}/3 CF rules found ({cf_score:.2f} pts)")
            total_score += cf_score
        else:
            print(f"FAIL: Component 4 — No conditional formatting rules found for course types")
            print(f"  Expected: 3 expression rules covering Core Requirement/Elective/Free Elective")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
