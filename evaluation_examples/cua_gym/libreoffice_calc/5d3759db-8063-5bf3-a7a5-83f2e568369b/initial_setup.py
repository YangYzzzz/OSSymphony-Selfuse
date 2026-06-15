"""
Initial Setup: Apply custom number format to Target Achievement columns
Task ID: calc_gsd_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Achievement'

    # Headers
    headers = ['Sales Rep', 'Q1 Target', 'Q1 Actual', 'Q1 Achievement%',
               'Q2 Target', 'Q2 Actual', 'Q2 Achievement%']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 25 sales rep records with realistic data
    sales_reps = [
        ('Sarah Chen', 85000, 96050, 1.13, 90000, 81000, 0.90),
        ('Marcus Johnson', 72000, 68400, 0.95, 78000, 85020, 1.09),
        ('Elena Rodriguez', 95000, 106400, 1.12, 92000, 71760, 0.78),
        ('David Kim', 68000, 91800, 1.35, 75000, 82500, 1.10),
        ('Priya Patel', 110000, 104500, 0.95, 105000, 115500, 1.10),
        ('James O\'Brien', 78000, 82680, 1.06, 82000, 73800, 0.90),
        ('Aisha Mohammed', 92000, 87400, 0.95, 88000, 97680, 1.11),
        ('Carlos Gutierrez', 65000, 72150, 1.11, 70000, 63000, 0.90),
        ('Lisa Thompson', 88000, 75680, 0.86, 91000, 100100, 1.10),
        ('Wei Zhang', 103000, 113300, 1.10, 99000, 89100, 0.90),
        ('Rachel Green', 76000, 79800, 1.05, 80000, 72000, 0.90),
        ('Omar Hassan', 84000, 71400, 0.85, 87000, 95700, 1.10),
        ('Sophie Martin', 97000, 106700, 1.10, 94000, 84600, 0.90),
        ('Tyler Washington', 71000, 63900, 0.90, 74000, 81400, 1.10),
        ('Yuki Tanaka', 89000, 97900, 1.10, 86000, 77400, 0.90),
        ('Natasha Volkov', 82000, 90200, 1.10, 85000, 93500, 1.10),
        ('Daniel Foster', 66000, 56100, 0.85, 69000, 75900, 1.10),
        ('Maria Santos', 93000, 102300, 1.10, 96000, 86400, 0.90),
        ('Kevin Park', 77000, 73150, 0.95, 81000, 89100, 1.10),
        ('Amanda Collins', 101000, 111100, 1.10, 98000, 88200, 0.90),
        ('Ibrahim Diallo', 69000, 62100, 0.90, 72000, 79200, 1.10),
        ('Hannah Lewis', 86000, 94600, 1.10, 83000, 74700, 0.90),
        ('Ryan Mitchell', 74000, 66600, 0.90, 77000, 84700, 1.10),
        ('Fatima Al-Rashid', 91000, 100100, 1.10, 89000, 80100, 0.90),
        ('Christopher Davis', 80000, 88000, 1.10, 84000, 92400, 1.10),
    ]

    for r, (name, q1t, q1a, q1pct, q2t, q2a, q2pct) in enumerate(sales_reps, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=q1t)
        ws.cell(row=r, column=3, value=q1a)
        ws.cell(row=r, column=4, value=q1pct)   # Plain number, no formatting
        ws.cell(row=r, column=5, value=q2t)
        ws.cell(row=r, column=6, value=q2a)
        ws.cell(row=r, column=7, value=q2pct)   # Plain number, no formatting

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
