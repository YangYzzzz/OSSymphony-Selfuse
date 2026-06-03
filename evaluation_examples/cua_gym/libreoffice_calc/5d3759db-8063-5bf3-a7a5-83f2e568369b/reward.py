"""
Reward Script: Apply custom number format to Achievement columns with color coding
Task ID: calc_gsd_032
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): D2:D26 have custom number format with percentage + color conditions
  Component 2 (0.35): G2:G26 have custom number format with percentage + color conditions
  Component 3 (0.15): Format includes [Green] for >=1 and [Red] for <1
  Component 4 (0.15): Underlying decimal values unchanged
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_032'

# Expected values from the initial/golden file (identical in both)
EXPECTED_D_VALUES = {
    2: 1.13, 3: 0.95, 4: 1.12, 5: 1.35, 6: 0.95, 7: 1.06, 8: 0.95,
    9: 1.11, 10: 0.86, 11: 1.1, 12: 1.05, 13: 0.85, 14: 1.1, 15: 0.9,
    16: 1.1, 17: 1.1, 18: 0.85, 19: 1.1, 20: 0.95, 21: 1.1, 22: 0.9,
    23: 1.1, 24: 0.9, 25: 1.1, 26: 1.1,
}

EXPECTED_G_VALUES = {
    2: 0.9, 3: 1.09, 4: 0.78, 5: 1.1, 6: 1.1, 7: 0.9, 8: 1.11,
    9: 0.9, 10: 1.1, 11: 0.9, 12: 0.9, 13: 1.1, 14: 0.9, 15: 1.1,
    16: 0.9, 17: 1.1, 18: 1.1, 19: 0.9, 20: 1.1, 21: 0.9, 22: 1.1,
    23: 0.9, 24: 1.1, 25: 0.9, 26: 1.1,
}


def is_custom_color_percent_format(fmt):
    """
    Check if a number format string is a custom format with:
    - Color conditions (green for high, red for low)
    - Percentage display (0.0% or similar)
    Returns (has_color_conditions, has_green, has_red, has_percent)
    """
    if fmt is None or fmt == 'General':
        return False, False, False, False

    fmt_upper = fmt.upper()
    has_green = '[GREEN]' in fmt_upper
    has_red = '[RED]' in fmt_upper
    # Check for percentage format pattern (digits with %)
    has_percent = '%' in fmt
    # Check for conditional syntax like [>=1] or [<1]
    has_condition = bool(re.search(r'\[([><=]+\s*\d)', fmt))

    has_color_conditions = (has_green or has_red) and has_condition
    return has_color_conditions, has_green, has_red, has_percent


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Achievement sheet exists
    if 'Achievement' not in wb.sheetnames:
        print("FAIL: 'Achievement' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Achievement']

    # Component 1: D2:D26 have custom number format with percentage + color conditions (0.35 points)
    try:
        d_custom_count = 0
        d_total = 25  # rows 2-26
        for r in range(2, 27):
            cell = ws.cell(row=r, column=4)  # column D
            fmt = cell.number_format
            has_color_cond, has_green, has_red, has_pct = is_custom_color_percent_format(fmt)
            if has_color_cond and has_pct:
                d_custom_count += 1

        if d_custom_count == d_total:
            print(f"PASS: Component 1 - All {d_total} cells in D2:D26 have custom color-conditional percentage format (0.35 pts)")
            total_score += 0.35
        elif d_custom_count > 0:
            partial = 0.35 * (d_custom_count / d_total)
            print(f"PARTIAL: Component 1 - {d_custom_count}/{d_total} cells in D2:D26 have custom format ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No cells in D2:D26 have custom color-conditional percentage format")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: G2:G26 have custom number format with percentage + color conditions (0.35 points)
    try:
        g_custom_count = 0
        g_total = 25  # rows 2-26
        for r in range(2, 27):
            cell = ws.cell(row=r, column=7)  # column G
            fmt = cell.number_format
            has_color_cond, has_green, has_red, has_pct = is_custom_color_percent_format(fmt)
            if has_color_cond and has_pct:
                g_custom_count += 1

        if g_custom_count == g_total:
            print(f"PASS: Component 2 - All {g_total} cells in G2:G26 have custom color-conditional percentage format (0.35 pts)")
            total_score += 0.35
        elif g_custom_count > 0:
            partial = 0.35 * (g_custom_count / g_total)
            print(f"PARTIAL: Component 2 - {g_custom_count}/{g_total} cells in G2:G26 have custom format ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No cells in G2:G26 have custom color-conditional percentage format")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Format includes both [Green] for >=1 AND [Red] for <1 (0.15 points)
    try:
        # Sample a few cells from D and G to check format details
        sample_cells = [ws.cell(row=2, column=4), ws.cell(row=2, column=7),
                        ws.cell(row=10, column=4), ws.cell(row=10, column=7)]
        green_found = 0
        red_found = 0
        for cell in sample_cells:
            fmt = cell.number_format
            if fmt and fmt != 'General':
                fmt_upper = fmt.upper()
                if '[GREEN]' in fmt_upper:
                    green_found += 1
                if '[RED]' in fmt_upper:
                    red_found += 1

        if green_found >= 2 and red_found >= 2:
            print(f"PASS: Component 3 - Format includes both [Green] and [Red] color codes (0.15 pts)")
            total_score += 0.15
        elif green_found >= 1 or red_found >= 1:
            print(f"PARTIAL: Component 3 - Only partial color coding found (green:{green_found}, red:{red_found}) (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 3 - No [Green]/[Red] color codes found in format strings")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Underlying decimal values unchanged (0.15 points)
    # This is a data integrity check anchored to format change — only scores if
    # format HAS been changed (so it won't score on initial_env where format is General)
    try:
        # First check: at least some format change exists (gate)
        format_change_count = sum(
            1 for r in range(2, 27)
            if ws.cell(row=r, column=4).number_format != 'General'
            or ws.cell(row=r, column=7).number_format != 'General'
        )

        if format_change_count == 0:
            print(f"FAIL: Component 4 - No format changes detected, so data integrity check not applicable")
        else:
            mismatch_count = 0
            for r in range(2, 27):
                d_val = ws.cell(row=r, column=4).value
                g_val = ws.cell(row=r, column=7).value
                if d_val is None or abs(float(d_val) - EXPECTED_D_VALUES[r]) > 0.011:
                    mismatch_count += 1
                if g_val is None or abs(float(g_val) - EXPECTED_G_VALUES[r]) > 0.011:
                    mismatch_count += 1

            if mismatch_count == 0:
                print(f"PASS: Component 4 - All underlying values preserved correctly (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 4 - {mismatch_count} value mismatches detected")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
