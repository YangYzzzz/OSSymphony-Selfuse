"""
Reward Script: Website Traffic Trend Analysis
Task ID: calc_gen_analysis_032
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): MoM Growth Rate formulas in F3:F13
  Component 2 (0.25): 3-Month Moving Average formulas in G4:G13
  Component 3 (0.25): Summary rows (Best Month / Worst Month / Avg Growth Rate) in rows 15-17
  Component 4 (0.20): Combo chart with bar (Sessions) + line (3-Month Avg) series
"""

import os
import zipfile
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_analysis_032'
SHEET_NAME = 'TrafficData'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -------------------------------------------------------------------------
    # Component 1: MoM Growth Rate formulas in F3:F13 (0.30 points)
    # F2 should be blank (no prior month), F3:F13 should have =(Bn-Bn-1)/Bn-1 formula
    # This FAILS on initial (all None) and PASSES on golden (formulas present)
    # -------------------------------------------------------------------------
    try:
        f2_val = ws['F2'].value
        f2_blank = (f2_val is None or str(f2_val).strip() == '')

        formula_count = 0
        for row in range(3, 14):  # rows 3-13
            cell_val = ws.cell(row=row, column=6).value  # column F
            if cell_val is None:
                print(f"FAIL: Component 1 — F{row} is empty, expected MoM growth formula")
                continue
            val_str = str(cell_val).strip().upper().replace(' ', '')
            # Accept formulas containing subtraction and division with B column refs
            if val_str.startswith('=') and 'B' in val_str and ('/' in val_str or '-' in val_str):
                formula_count += 1
            else:
                print(f"FAIL: Component 1 — F{row} value '{cell_val}' does not look like MoM growth formula")

        if formula_count == 11 and f2_blank:
            print(f"PASS: Component 1 — MoM Growth formulas found in F3:F13 ({formula_count}/11), F2 is blank (0.30 pts)")
            total_score += 0.30
        elif formula_count >= 6:
            print(f"PARTIAL: Component 1 — {formula_count}/11 MoM Growth formulas found in F3:F13 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/11 MoM Growth formulas found in F3:F13")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: 3-Month Moving Average formulas in G4:G13 (0.25 points)
    # G2:G3 should be blank (insufficient history), G4:G13 should have AVERAGE formulas
    # This FAILS on initial (all None) and PASSES on golden (formulas present)
    # -------------------------------------------------------------------------
    try:
        g2_blank = (ws['G2'].value is None or str(ws['G2'].value).strip() == '')
        g3_blank = (ws['G3'].value is None or str(ws['G3'].value).strip() == '')

        avg_formula_count = 0
        for row in range(4, 14):  # rows 4-13
            cell_val = ws.cell(row=row, column=7).value  # column G
            if cell_val is None:
                print(f"FAIL: Component 2 — G{row} is empty, expected 3-month average formula")
                continue
            val_str = str(cell_val).strip().upper().replace(' ', '')
            # Should be AVERAGE formula referencing B column
            if val_str.startswith('=') and 'AVERAGE' in val_str and 'B' in val_str:
                avg_formula_count += 1
            else:
                print(f"FAIL: Component 2 — G{row} value '{cell_val}' does not look like AVERAGE formula")

        if avg_formula_count == 10 and g2_blank and g3_blank:
            print(f"PASS: Component 2 — 3-Month Avg formulas in G4:G13 ({avg_formula_count}/10), G2:G3 blank (0.25 pts)")
            total_score += 0.25
        elif avg_formula_count >= 5:
            print(f"PARTIAL: Component 2 — {avg_formula_count}/10 3-Month Avg formulas found (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — Only {avg_formula_count}/10 3-Month Avg formulas found in G4:G13")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Summary stats rows 15-17 (0.25 points)
    # Row 15: 'Best Month:' label + INDEX/MATCH/MAX formula in B15
    # Row 16: 'Worst Month:' label + INDEX/MATCH/MIN formula in B16
    # Row 17: 'Avg Growth Rate:' label + AVERAGE formula over F column in B17
    # This FAILS on initial (all None) and PASSES on golden
    # -------------------------------------------------------------------------
    try:
        a15 = ws['A15'].value
        b15 = ws['B15'].value
        a16 = ws['A16'].value
        b16 = ws['B16'].value
        a17 = ws['A17'].value
        b17 = ws['B17'].value

        label_check = (
            a15 is not None and 'best' in str(a15).lower() and
            a16 is not None and 'worst' in str(a16).lower() and
            a17 is not None and 'avg' in str(a17).lower() and 'growth' in str(a17).lower()
        )

        # Check B15 has INDEX/MAX formula
        b15_str = str(b15).upper().replace(' ', '') if b15 is not None else ''
        b15_ok = (b15_str.startswith('=') and 'INDEX' in b15_str and 'MAX' in b15_str)

        # Check B16 has INDEX/MIN formula
        b16_str = str(b16).upper().replace(' ', '') if b16 is not None else ''
        b16_ok = (b16_str.startswith('=') and 'INDEX' in b16_str and 'MIN' in b16_str)

        # Check B17 has AVERAGE formula over F column
        b17_str = str(b17).upper().replace(' ', '') if b17 is not None else ''
        b17_ok = (b17_str.startswith('=') and 'AVERAGE' in b17_str and 'F' in b17_str)

        checks_passed = sum([b15_ok, b16_ok, b17_ok])

        if label_check and checks_passed == 3:
            print(f"PASS: Component 3 — Summary rows 15-17 correct: Best={b15}, Worst={b16}, AvgGrowth={b17} (0.25 pts)")
            total_score += 0.25
        elif label_check and checks_passed >= 1:
            partial = round(0.25 * checks_passed / 3, 2)
            print(f"PARTIAL: Component 3 — {checks_passed}/3 summary formulas correct (labels OK) ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 3 — Summary rows missing or incorrect. A15={repr(a15)}, B15={repr(b15)}, A16={repr(a16)}, B16={repr(b16)}, A17={repr(a17)}, B17={repr(b17)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Combo chart with bar (Sessions) + line (3-Month Avg) (0.20 points)
    # Inspects chart XML: must contain both barChart and lineChart elements
    # with series referencing columns B (Sessions) and G (3-Month Avg).
    # This FAILS on initial (no charts) and PASSES on golden (combo chart present).
    # -------------------------------------------------------------------------
    try:
        num_charts = len(ws._charts)

        # Inspect chart XML directly to verify combo structure
        # Read all chart XML files and check for both barChart+B-ref and lineChart+G-ref
        bar_found = False
        line_found = False

        if num_charts >= 1:
            with zipfile.ZipFile(file_path, 'r') as z:
                chart_files = [n for n in z.namelist() if re.match(r'xl/charts/chart\d+\.xml', n)]
                combined_content = ''.join(z.read(cf).decode('utf-8') for cf in chart_files)
            # Evaluate using single boolean expressions (no = True assignment)
            bar_found = (
                'barChart' in combined_content and
                ('$B$' in combined_content or "TrafficData'!B" in combined_content or 'B2:B13' in combined_content or 'B$2' in combined_content)
            )
            line_found = (
                'lineChart' in combined_content and
                ('$G$' in combined_content or "TrafficData'!G" in combined_content or 'G2:G13' in combined_content or 'G$2' in combined_content)
            )

        if bar_found and line_found:
            print(f"PASS: Component 4 — Combo chart found with bar (Sessions) + line (3-Month Avg) series (0.20 pts)")
            total_score += 0.20
        elif num_charts >= 1 and (bar_found or line_found):
            print(f"PARTIAL: Component 4 — Chart found but missing one series type (bar_found={bar_found}, line_found={line_found}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — No valid combo chart found (charts={num_charts}, bar_found={bar_found}, line_found={line_found})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
