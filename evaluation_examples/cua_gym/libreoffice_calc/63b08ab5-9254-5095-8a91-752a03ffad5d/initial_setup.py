"""
Initial Setup: Website traffic data with 12 months of sessions, page views, bounce rate,
avg duration. Columns F (MoM Growth %) and G (3-Month Avg) are intentionally empty.
Task ID: calc_gen_analysis_032
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_analysis_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'TrafficData'

    # --- Headers ---
    headers = ['Month', 'Sessions', 'Page Views', 'Bounce Rate', 'Avg Duration', 'MoM Growth %', '3-Month Avg']
    header_font = Font(bold=True, name='Calibri', size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Data: 12 months of realistic website traffic ---
    # Sessions: realistic monthly totals, Page Views ~3x Sessions,
    # Bounce Rate: 40-65%, Avg Duration: 1m30s - 4m in seconds
    # F and G columns intentionally left empty
    data = [
        ('January',   42310,  124580, 0.623, 148),
        ('February',  38920,  114270, 0.591, 162),
        ('March',     51840,  157320, 0.548, 195),
        ('April',     63750,  196420, 0.512, 218),
        ('May',       71280,  221460, 0.487, 234),
        ('June',      68540,  209870, 0.503, 227),
        ('July',      75320,  238750, 0.471, 251),
        ('August',    82190,  264340, 0.445, 267),
        ('September', 79640,  252870, 0.458, 259),
        ('October',   88430,  285620, 0.432, 281),
        ('November',  95780,  312470, 0.418, 298),
        ('December',  91250,  295340, 0.437, 272),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Month
        ws.cell(row=r, column=2, value=row_data[1])   # Sessions
        ws.cell(row=r, column=3, value=row_data[2])   # Page Views
        cell_bounce = ws.cell(row=r, column=4, value=row_data[3])   # Bounce Rate
        cell_bounce.number_format = '0.0%'
        ws.cell(row=r, column=5, value=row_data[4])   # Avg Duration (seconds)
        # Columns F (6) and G (7) intentionally left empty

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
