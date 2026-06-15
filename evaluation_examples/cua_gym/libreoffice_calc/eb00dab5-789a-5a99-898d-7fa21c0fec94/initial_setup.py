"""
Initial Setup: Set up a spreadsheet with data table and chart, print area covers entire used range
Task ID: calc_chart_print_area_chart_078
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_print_area_chart_078'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Report ---
    ws = wb.active
    ws.title = 'Report'

    # Headers in A1:C1
    headers = ['Category', 'Plan', 'Actual']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Data rows for Jan-Jul (A2:C8)
    data = [
        ['Jan', 120000, 115400],
        ['Feb', 135000, 138200],
        ['Mar', 142000, 140500],
        ['Apr', 128000, 131800],
        ['May', 155000, 149700],
        ['Jun', 162000, 165300],
        ['Jul', 148000, 152600],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    # Create a clustered column chart
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Monthly Plan vs Actual'
    chart.y_axis.title = 'Amount ($)'
    chart.x_axis.title = 'Month'
    chart.style = 10

    # Data reference: B1:C8 (with titles)
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=8)
    cats = Reference(ws, min_col=1, min_row=2, max_row=8)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    # Position chart in E1:L20 (anchor at E1, set size to cover ~L20)
    # Each column is ~64px wide, each row is ~20px tall
    # E1 to L20 means chart spans 8 columns (E-L) x 20 rows
    chart.width = 16    # approx width in cm to fill E:L
    chart.height = 10   # approx height in cm to fill rows 1-20

    ws.add_chart(chart, 'E1')

    # Set print area to A1:L20 (both data table and chart)
    ws.print_area = 'A1:L20'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: Report')
    print('  Data: A1:C8 (Category, Plan, Actual) with Jan-Jul rows')
    print('  Chart: Clustered column chart at E1')
    print('  Print area: A1:L20 (data + chart)')


create_initial()
