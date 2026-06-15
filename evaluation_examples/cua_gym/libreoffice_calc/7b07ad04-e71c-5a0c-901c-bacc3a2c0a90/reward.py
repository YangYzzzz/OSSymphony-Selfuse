"""
Reward Script: Process utility bills (PDFs) and enter them into monthly expense tracker in Calc
Task ID: osworld_multi_apps_doc_pdf_calc_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): 5 new March 2025 rows added with correct provider, category, and amount data
  Component 2 (0.3): All 5 March 2025 entries have correct amounts matching the PDFs
  Component 3 (0.2): Running total formula updated to cover March rows (Jan-Mar)
"""

import os
import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_doc_pdf_calc_006'

# Expected March 2025 data from the PDFs (per task_config.json context)
EXPECTED_MARCH_ROWS = [
    ("Pacific Gas & Electric", "Electric", "March 2025", 143.22),
    ("City Water Dept", "Water", "March 2025", 45.80),
    ("SoCalGas", "Gas", "March 2025", 89.15),
    ("Comcast", "Internet", "March 2025", 79.99),
    ("Verizon", "Phone", "March 2025", 65.00),
]

# Expected updated running total label in golden
EXPECTED_TOTAL_LABEL = "Running Total (Jan-Mar)"
# Running total formula should reference at least through D16 (after 5 new rows added)
# In golden: =SUM(D2:D16)
EXPECTED_TOTAL_FORMULA_COVERS = "D16"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active
    if ws is None:
        print("CRITICAL: No active sheet found")
        print("REWARD: 0.0")
        return 0.0

    # Collect all data rows (skip header row 1)
    all_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        all_rows.append(row)

    # Find all March 2025 rows in the spreadsheet
    march_rows = [r for r in all_rows if r and len(r) >= 3 and r[2] == "March 2025"]

    # Component 1: 5 new March 2025 rows added with correct providers and categories (0.5 points)
    try:
        # Build set of (provider, category, month) tuples found
        found_march_entries = set()
        for r in march_rows:
            if r[0] and r[1] and r[2]:
                found_march_entries.add((str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip()))

        expected_providers_cats = {
            (provider.strip(), category.strip(), month.strip())
            for provider, category, month, _ in EXPECTED_MARCH_ROWS
        }

        # Count how many of the expected entries are present
        matched_entries = expected_providers_cats & found_march_entries
        if len(matched_entries) == 5:
            print(f"PASS: Component 1 — All 5 March 2025 rows added with correct provider/category (0.5 pts)")
            total_score += 0.5
        elif len(matched_entries) >= 1:
            partial = round(0.1 * len(matched_entries), 1)
            print(f"PARTIAL: Component 1 — {len(matched_entries)}/5 March 2025 rows have correct provider/category ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No March 2025 rows with correct provider/category found (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 5 March 2025 amounts are correct (0.3 points)
    try:
        # Build dict from found march rows: (provider, category, month) -> amount
        found_amounts = {}
        for r in march_rows:
            if r[0] and r[1] and r[2] and r[3] is not None:
                key = (str(r[0]).strip(), str(r[1]).strip(), str(r[2]).strip())
                try:
                    found_amounts[key] = float(r[3])
                except (ValueError, TypeError):
                    pass

        correct_amounts = 0
        for provider, category, month, expected_amount in EXPECTED_MARCH_ROWS:
            key = (provider.strip(), category.strip(), month.strip())
            if key in found_amounts:
                actual = found_amounts[key]
                if abs(actual - expected_amount) <= 0.01:
                    correct_amounts += 1
                else:
                    print(f"  FAIL: Amount for {provider}/{month}: expected {expected_amount}, found {actual}")
            else:
                print(f"  FAIL: No matching row found for {provider}/{category}/{month}")

        if correct_amounts == 5:
            print(f"PASS: Component 2 — All 5 March 2025 amounts correct (0.3 pts)")
            total_score += 0.3
        elif correct_amounts >= 1:
            partial = round(0.06 * correct_amounts, 2)
            print(f"PARTIAL: Component 2 — {correct_amounts}/5 March 2025 amounts correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No March 2025 amounts are correct (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Running total label and formula updated to include March rows (0.2 points)
    try:
        # Find the running total row (row with "Running Total" in column C)
        total_row = None
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            c_val = row[2].value if len(row) >= 3 else None
            if c_val and isinstance(c_val, str) and "Running Total" in c_val:
                total_row = row
                break

        if total_row is None:
            print("FAIL: Component 3 — No 'Running Total' row found")
        else:
            label_val = total_row[2].value if len(total_row) >= 3 else None
            formula_val = total_row[3].value if len(total_row) >= 4 else None

            label_correct = label_val == EXPECTED_TOTAL_LABEL
            # Formula should reference at least D16 (covers March rows rows 12-16)
            formula_correct = (
                formula_val is not None
                and isinstance(formula_val, str)
                and "SUM" in formula_val.upper()
                and EXPECTED_TOTAL_FORMULA_COVERS.upper() in formula_val.upper()
            )

            if label_correct and formula_correct:
                print(f"PASS: Component 3 — Running total label is '{label_val}' and formula is '{formula_val}' (0.2 pts)")
                total_score += 0.2
            elif label_correct:
                print(f"PARTIAL: Component 3 — Label correct ('{label_val}') but formula incorrect: '{formula_val}' (0.1 pts)")
                total_score += 0.1
            elif formula_correct:
                print(f"PARTIAL: Component 3 — Formula correct ('{formula_val}') but label incorrect: '{label_val}' (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 3 — Label: '{label_val}', Formula: '{formula_val}' (expected label='Running Total (Jan-Mar)', formula covers D16) (0 pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against the canonical artifact path on VM
file_path = f'{WORKDIR}/utility_tracker.xlsx'
if not os.path.exists(file_path):
    # Fallback to .ods if .xlsx not found
    file_path_ods = f'{WORKDIR}/utility_tracker.ods'
    if os.path.exists(file_path_ods):
        print(f"NOTE: .xlsx not found, trying .ods: {file_path_ods}")
        file_path = file_path_ods
    else:
        print(f"File not found: {file_path}")
        print("REWARD: 0.0")
        exit(0)

verify_task(file_path)
