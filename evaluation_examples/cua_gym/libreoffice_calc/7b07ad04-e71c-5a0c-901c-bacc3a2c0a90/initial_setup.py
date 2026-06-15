"""
Initial Setup: Process utility bills PDF folder and enter into monthly expense tracker
Task ID: osworld_multi_apps_doc_pdf_calc_006
Domain: libreoffice_calc (multi-app: PDF files + Calc spreadsheet)

Creates:
  - /home/user/Desktop/utilities/ folder with 5 PDF utility bills
  - /home/user/Desktop/utility_tracker.ods (prior months data, NO March 2025 rows)
  - Opens utility_tracker.ods in LibreOffice Calc for the agent
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_doc_pdf_calc_006'
DESKTOP = f'{WORKDIR}/Desktop'
UTILITIES_DIR = f'{DESKTOP}/utilities'
TRACKER_ODS = f'{DESKTOP}/utility_tracker.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_pdf_bill(filepath, provider, utility_type, month_str, amount, account_num, service_addr):
    """Create a simple text-based PDF utility bill using fpdf2."""
    try:
        from fpdf import FPDF

        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Helvetica", size=12)

        # Header
        pdf.set_font("Helvetica", style="B", size=16)
        pdf.cell(0, 10, provider, ln=True, align="C")
        pdf.set_font("Helvetica", size=10)
        pdf.cell(0, 6, "Utility Bill Statement", ln=True, align="C")
        pdf.ln(5)

        # Horizontal line
        pdf.set_draw_color(100, 100, 100)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(5)

        # Account info
        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(60, 8, "Account Number:", ln=False)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, account_num, ln=True)

        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(60, 8, "Service Address:", ln=False)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, service_addr, ln=True)

        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(60, 8, "Utility Type:", ln=False)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, utility_type, ln=True)

        pdf.ln(5)
        pdf.set_draw_color(100, 100, 100)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(5)

        # Billing period
        pdf.set_font("Helvetica", style="B", size=12)
        pdf.cell(0, 8, "BILLING SUMMARY", ln=True)
        pdf.ln(3)

        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(80, 8, "Billing Month:", ln=False)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, month_str, ln=True)

        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(80, 8, "Previous Balance:", ln=False)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, "$0.00", ln=True)

        pdf.set_font("Helvetica", style="B", size=11)
        pdf.cell(80, 8, "Current Charges:", ln=False)
        pdf.set_font("Helvetica", size=11)
        pdf.cell(0, 8, f"${amount:.2f}", ln=True)

        pdf.ln(5)
        pdf.set_draw_color(0, 0, 0)
        pdf.line(10, pdf.get_y(), 200, pdf.get_y())
        pdf.ln(3)

        pdf.set_font("Helvetica", style="B", size=13)
        pdf.cell(80, 10, "AMOUNT DUE:", ln=False)
        pdf.set_font("Helvetica", style="B", size=13)
        pdf.cell(0, 10, f"${amount:.2f}", ln=True)

        pdf.ln(5)
        pdf.set_font("Helvetica", size=9)
        pdf.cell(0, 6, "Payment due within 30 days of billing date.", ln=True)
        pdf.cell(0, 6, "For questions call 1-800-555-0100 or visit our website.", ln=True)

        pdf.output(filepath)
        print(f"  Created PDF: {filepath}")
    except ImportError:
        # Fallback: create a text file disguised as PDF-like content
        # Actually write a minimal valid PDF manually
        content = f"""%PDF-1.4
1 0 obj
<< /Type /Catalog /Pages 2 0 R >>
endobj
2 0 obj
<< /Type /Pages /Kids [3 0 R] /Count 1 >>
endobj
3 0 obj
<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792]
   /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>
endobj
4 0 obj
<< /Length 400 >>
stream
BT
/F1 16 Tf
50 740 Td
({provider}) Tj
/F1 11 Tf
0 -25 Td
(Utility Bill Statement) Tj
0 -30 Td
(Billing Month: {month_str}) Tj
0 -20 Td
(Provider: {provider}) Tj
0 -20 Td
(Utility Type: {utility_type}) Tj
0 -20 Td
(Account: {account_num}) Tj
0 -30 Td
/F1 13 Tf
(AMOUNT DUE: ${amount:.2f}) Tj
ET
endstream
endobj
5 0 obj
<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>
endobj
xref
0 6
0000000000 65535 f
0000000009 00000 n
0000000058 00000 n
0000000115 00000 n
0000000266 00000 n
0000000716 00000 n
trailer
<< /Size 6 /Root 1 0 R >>
startxref
800
%%EOF"""
        with open(filepath, 'w') as f:
            f.write(content)
        print(f"  Created PDF (fallback): {filepath}")


def create_tracker_ods():
    """Create utility_tracker.ods with prior months data (Jan and Feb 2025), NO March 2025."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Utility Expenses"

    # --- Headers ---
    headers = ["Provider", "Category", "Month", "Amount"]
    header_font = Font(bold=True, color="FFFFFFFF", size=12)
    header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="FF000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    ws.row_dimensions[1].height = 22

    # --- Prior months data: January 2025 and February 2025 ---
    # January 2025 data
    jan_data = [
        ["Pacific Gas & Electric", "Electric", "January 2025", 128.45],
        ["City Water Dept",        "Water",    "January 2025",  42.30],
        ["SoCalGas",               "Gas",      "January 2025",  97.60],
        ["Comcast",                "Internet",  "January 2025",  79.99],
        ["Verizon",                "Phone",    "January 2025",  65.00],
    ]

    # February 2025 data
    feb_data = [
        ["Pacific Gas & Electric", "Electric", "February 2025", 135.18],
        ["City Water Dept",        "Water",    "February 2025",  44.55],
        ["SoCalGas",               "Gas",      "February 2025",  91.30],
        ["Comcast",                "Internet",  "February 2025",  79.99],
        ["Verizon",                "Phone",    "February 2025",  65.00],
    ]

    all_data = jan_data + feb_data
    data_fill_even = PatternFill(start_color="FFDCE6F1", end_color="FFDCE6F1", fill_type="solid")
    data_fill_odd = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row_data in enumerate(all_data, 2):
        fill = data_fill_even if r_idx % 2 == 0 else data_fill_odd
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = data_border
            cell.fill = fill
            if c_idx == 4:  # Amount column
                cell.number_format = '$#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            elif c_idx == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Number of data rows written (Jan + Feb = 10 rows, rows 2-11)
    last_data_row = 1 + len(all_data)  # row 11

    # Running total label and formula for existing data
    # Place a "Subtotal" row after the current data
    subtotal_row = last_data_row + 1  # row 12
    subtotal_label_cell = ws.cell(row=subtotal_row, column=3, value="Running Total (Jan-Feb)")
    subtotal_label_cell.font = Font(bold=True, italic=True)
    subtotal_label_cell.alignment = Alignment(horizontal="right")

    subtotal_value_cell = ws.cell(row=subtotal_row, column=4,
                                   value=f"=SUM(D2:D{last_data_row})")
    subtotal_value_cell.font = Font(bold=True)
    subtotal_value_cell.number_format = '$#,##0.00'
    subtotal_value_cell.alignment = Alignment(horizontal="right")

    # Column widths
    ws.column_dimensions["A"].width = 28  # Provider
    ws.column_dimensions["B"].width = 12  # Category
    ws.column_dimensions["C"].width = 18  # Month
    ws.column_dimensions["D"].width = 14  # Amount

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save as .xlsx (openpyxl native format; LibreOffice Calc opens .xlsx)
    wb.save(TRACKER_ODS)

    print(f"Tracker created: {TRACKER_ODS}")
    return last_data_row


def create_initial():
    # Ensure Desktop exists
    os.makedirs(DESKTOP, exist_ok=True)

    # --- Create utilities/ folder with 5 PDFs ---
    os.makedirs(UTILITIES_DIR, exist_ok=True)
    print(f"Created utilities folder: {UTILITIES_DIR}")

    bills = [
        {
            "filename": "electric_bill.pdf",
            "provider": "Pacific Gas & Electric",
            "utility_type": "Electricity",
            "month_str": "March 2025",
            "amount": 143.22,
            "account_num": "ACC-4872-EL",
            "service_addr": "1428 Elm Street, Oakland, CA 94601",
        },
        {
            "filename": "water_bill.pdf",
            "provider": "City Water Dept",
            "utility_type": "Water",
            "month_str": "March 2025",
            "amount": 45.80,
            "account_num": "ACC-2031-WA",
            "service_addr": "1428 Elm Street, Oakland, CA 94601",
        },
        {
            "filename": "gas_bill.pdf",
            "provider": "SoCalGas",
            "utility_type": "Natural Gas",
            "month_str": "March 2025",
            "amount": 89.15,
            "account_num": "ACC-5519-GS",
            "service_addr": "1428 Elm Street, Oakland, CA 94601",
        },
        {
            "filename": "internet_bill.pdf",
            "provider": "Comcast",
            "utility_type": "Internet",
            "month_str": "March 2025",
            "amount": 79.99,
            "account_num": "ACC-7743-IN",
            "service_addr": "1428 Elm Street, Oakland, CA 94601",
        },
        {
            "filename": "phone_bill.pdf",
            "provider": "Verizon",
            "utility_type": "Phone",
            "month_str": "March 2025",
            "amount": 65.00,
            "account_num": "ACC-9912-PH",
            "service_addr": "1428 Elm Street, Oakland, CA 94601",
        },
    ]

    for bill in bills:
        pdf_path = os.path.join(UTILITIES_DIR, bill["filename"])
        create_pdf_bill(
            pdf_path,
            bill["provider"],
            bill["utility_type"],
            bill["month_str"],
            bill["amount"],
            bill["account_num"],
            bill["service_addr"],
        )

    # --- Create utility_tracker spreadsheet ---
    create_tracker_ods()

    # --- GUI-ready startup ---
    # Open the tracker file in LibreOffice Calc
    # Also open the file manager to show the utilities folder with PDFs
    launch_gui(f'libreoffice --calc "{TRACKER_ODS}"', delay_sec=2.5)
    launch_gui(f'nautilus "{UTILITIES_DIR}"', delay_sec=1.5)
    print('GUI_READY: Launched LibreOffice Calc with utility_tracker.ods and Nautilus with utilities/ folder')


create_initial()
