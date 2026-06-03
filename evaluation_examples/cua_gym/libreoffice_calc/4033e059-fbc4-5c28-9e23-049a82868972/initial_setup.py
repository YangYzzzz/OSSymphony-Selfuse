"""
Initial Setup: Configure page centering for certificate spreadsheet
Task ID: calc_gfl_087
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Certificate ---
    ws = wb.active
    ws.title = 'Certificate'

    # Certificate of Completion layout in A1:D10
    # Row 1: Merged title
    ws.merge_cells('A1:D1')
    ws['A1'] = 'CERTIFICATE OF COMPLETION'
    ws['A1'].font = Font(name='Arial', size=18, bold=True, color='1F4E79')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 2: Subtitle
    ws.merge_cells('A2:D2')
    ws['A2'] = 'This certificate is proudly presented to'
    ws['A2'].font = Font(name='Arial', size=11, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Row 3: Recipient name
    ws.merge_cells('A3:D3')
    ws['A3'] = 'Elena Vasquez'
    ws['A3'].font = Font(name='Arial', size=16, bold=True, color='2E75B6')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # Row 4: Description
    ws.merge_cells('A4:D4')
    ws['A4'] = 'for successful completion of the following program:'
    ws['A4'].font = Font(name='Arial', size=10)
    ws['A4'].alignment = Alignment(horizontal='center')

    # Row 5: Program name
    ws.merge_cells('A5:D5')
    ws['A5'] = 'Advanced Data Analytics & Business Intelligence'
    ws['A5'].font = Font(name='Arial', size=13, bold=True)
    ws['A5'].alignment = Alignment(horizontal='center')

    # Row 6: Blank spacer
    ws.merge_cells('A6:D6')
    ws['A6'] = ''

    # Row 7: Details header row
    ws['A7'] = 'Date Issued'
    ws['B7'] = 'Duration'
    ws['C7'] = 'Grade'
    ws['D7'] = 'Instructor'
    header_font = Font(name='Arial', size=10, bold=True)
    header_fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')
    for col in range(1, 5):
        cell = ws.cell(row=7, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Row 8: Details data
    ws['A8'] = '2025-09-15'
    ws['B8'] = '120 Hours'
    ws['C8'] = 'Distinction'
    ws['D8'] = 'Dr. Michael Torres'
    for col in range(1, 5):
        ws.cell(row=8, column=col).alignment = Alignment(horizontal='center')
        ws.cell(row=8, column=col).font = Font(name='Arial', size=10)

    # Row 9: Blank spacer
    ws.merge_cells('A9:D9')
    ws['A9'] = ''

    # Row 10: Organization
    ws.merge_cells('A10:D10')
    ws['A10'] = 'Pacific Northwest Institute of Technology'
    ws['A10'].font = Font(name='Arial', size=11, bold=True, color='1F4E79')
    ws['A10'].alignment = Alignment(horizontal='center')

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 22

    # Set row heights
    ws.row_dimensions[1].height = 35
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[5].height = 25

    # Page setup: default (NOT centered) - content prints top-left
    # Explicitly ensure centering is NOT set
    ws.print_options.horizontalCentered = False
    ws.print_options.verticalCentered = False

    # Set print area
    ws.print_area = 'A1:D10'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
