"""
Initial Setup: Count orders between $500 and $2,000 using COUNTIFS
Task ID: calc_fmb_countifs_range_069
Domain: libreoffice_calc
"""

import os
import random
import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_countifs_range_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Order Log'

    # --- Row 1: Headers ---
    ws['A1'] = 'Order ID'
    ws['B1'] = 'Customer'
    ws['C1'] = 'Order Value'
    ws['D1'] = 'Date'
    ws['E1'] = 'Status'
    ws['F1'] = 'Count'

    # --- E2: Label as specified in context ---
    ws['E2'] = '$500-$2000 Orders'
    # F2 is the target cell — MUST remain empty in initial

    # --- Customer names ---
    first_names = [
        'James', 'Emma', 'Liam', 'Olivia', 'Noah', 'Ava', 'William', 'Sophia',
        'Benjamin', 'Isabella', 'Mason', 'Mia', 'Ethan', 'Charlotte', 'Alexander',
        'Amelia', 'Henry', 'Harper', 'Sebastian', 'Evelyn', 'Michael', 'Abigail',
        'Daniel', 'Emily', 'Matthew', 'Elizabeth', 'Jackson', 'Mila', 'David',
        'Ella', 'Joseph', 'Avery', 'Samuel', 'Sofia', 'Carter', 'Camila',
        'Owen', 'Aria', 'Wyatt', 'Scarlett', 'John', 'Victoria', 'Jack',
        'Madison', 'Luke', 'Luna', 'Jayden', 'Grace', 'Dylan', 'Chloe'
    ]
    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
        'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
        'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
        'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
        'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
        'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green',
        'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts'
    ]

    statuses = ['Completed', 'Pending', 'Shipped', 'Cancelled', 'Processing']

    # We need exactly 187 orders in [500, 2000] range
    # Total orders: 500
    # In-range: 187
    # Out-of-range: 313 (either < 500 or > 2000)

    # Generate in-range values: 187 values between 500 and 2000 (inclusive)
    in_range_values = []
    for _ in range(187):
        val = round(random.uniform(500.00, 2000.00), 2)
        in_range_values.append(val)

    # Generate out-of-range values: 313 values either < 500 or > 2000
    # Mix: 150 below $500 (15-499.99), 163 above $2000 (2000.01-9800)
    out_range_values = []
    for _ in range(150):
        val = round(random.uniform(15.00, 499.99), 2)
        out_range_values.append(val)
    for _ in range(163):
        val = round(random.uniform(2000.01, 9800.00), 2)
        out_range_values.append(val)

    # Combine all 500 values and shuffle
    all_values = in_range_values + out_range_values
    random.shuffle(all_values)

    # Generate base date
    base_date = date(2024, 1, 1)

    # Populate rows 2-501
    for i, order_val in enumerate(all_values):
        row = i + 2
        order_id = f'ORD-{10000 + i}'
        customer = f'{random.choice(first_names)} {random.choice(last_names)}'
        order_date = base_date + timedelta(days=random.randint(0, 364))
        status = random.choice(statuses)

        ws.cell(row=row, column=1, value=order_id)
        ws.cell(row=row, column=2, value=customer)
        ws.cell(row=row, column=3, value=order_val)
        ws.cell(row=row, column=4, value=order_date.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=5, value=status if row != 2 else '$500-$2000 Orders')
        ws.cell(row=row, column=6, value=None)  # F column empty (target)

    # Fix E2 to be the label (row=2, col=5 already set above, but we need to override)
    ws['E2'] = '$500-$2000 Orders'
    # F2 must be empty
    ws['F2'] = None

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify: count orders in range
    count_in_range = sum(1 for v in all_values if 500 <= v <= 2000)
    print(f'Verification: Orders in $500-$2000 range: {count_in_range} (expected 187)')

create_initial()
