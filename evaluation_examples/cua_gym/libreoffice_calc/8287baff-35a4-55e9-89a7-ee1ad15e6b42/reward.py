"""
Reward Script: Count orders with values between $500 and $2,000 using COUNTIFS
Task ID: calc_fmb_countifs_range_069
Domain: libreoffice_calc
Scoring:
  Component 1: Cell F2 contains a COUNTIFS formula          (0.4 points)
  Component 2: Formula uses correct range C2:C501 with >=500 and <=2000 criteria  (0.4 points)
  Component 3: Formula result (187) verified by direct computation on data        (0.2 points)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_countifs_range_069'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Task: Count orders where order value is between $500 and $2,000 (inclusive).
          Order values are in column C (rows 2-501). Put count in cell F2.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Verify 'Order Log' sheet exists
    if 'Order Log' not in wb.sheetnames:
        print("CRITICAL: 'Order Log' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Order Log']

    # Component 1: Cell F2 contains a COUNTIFS formula (0.4 points)
    # The task requires placing a COUNTIFS formula in F2. This fails on
    # the initial file (F2 is None) and passes on the golden file.
    try:
        f2_value = ws['F2'].value
        if f2_value is not None and isinstance(f2_value, str) and 'COUNTIFS' in f2_value.upper():
            print(f"PASS: Component 1 — F2 contains a COUNTIFS formula: {repr(f2_value)} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — F2 should contain COUNTIFS formula, found: {repr(f2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not read F2: {e}")

    # Component 2: Formula uses correct range C2:C501 with bounds >=500 and <=2000 (0.4 points)
    # The formula must reference the correct data range (C2:C501) and use
    # the correct criteria (>=500 and <=2000). This fails on the initial
    # file and passes on the golden file.
    try:
        f2_value = ws['F2'].value
        if f2_value is not None and isinstance(f2_value, str):
            formula_upper = f2_value.upper().replace(' ', '')
            # Check that the formula references C2:C501 (the correct data range)
            has_correct_range = 'C2:C501' in formula_upper
            # Check for lower bound: >=500 (in various forms: >=500, >=&500, ">=500", ">="&500)
            has_lower_bound = bool(re.search(r'>=.{0,3}500', formula_upper))
            # Check for upper bound: <=2000 (in various forms)
            has_upper_bound = bool(re.search(r'<=.{0,3}2000', formula_upper))

            if has_correct_range and has_lower_bound and has_upper_bound:
                print(f"PASS: Component 2 — Formula references C2:C501 with >=500 and <=2000 criteria (0.4 pts)")
                total_score += 0.4
            else:
                issues = []
                if not has_correct_range:
                    issues.append("missing or wrong range (expected C2:C501)")
                if not has_lower_bound:
                    issues.append("missing lower bound >=500")
                if not has_upper_bound:
                    issues.append("missing upper bound <=2000")
                print(f"FAIL: Component 2 — Formula criteria issues: {', '.join(issues)}")
        else:
            print(f"FAIL: Component 2 — F2 is empty or not a formula, cannot check criteria")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not verify formula criteria: {e}")

    # Component 3: Verify the count result (187) by computing directly on the data (0.2 points)
    # We independently compute the COUNTIFS result by reading C2:C501 values
    # and counting those between 500 and 2000 inclusive. If the formula is
    # present AND data supports 187 orders, we award this bonus component.
    try:
        computed_count = 0
        for row in range(2, 502):
            val = ws.cell(row=row, column=3).value  # Column C = order values
            if val is not None:
                try:
                    v = float(val)
                    if 500 <= v <= 2000:
                        computed_count += 1
                except (ValueError, TypeError):
                    pass

        # The formula in F2 must be present, and the underlying data must yield 187
        f2_value = ws['F2'].value
        formula_present = (f2_value is not None and isinstance(f2_value, str) and
                           'COUNTIFS' in f2_value.upper())

        if formula_present and computed_count == 187:
            print(f"PASS: Component 3 — Data yields {computed_count} orders in $500-$2000 range, formula present (0.2 pts)")
            total_score += 0.2
        elif not formula_present:
            print(f"FAIL: Component 3 — Formula not present in F2, cannot verify count")
        else:
            print(f"FAIL: Component 3 — Expected 187 orders in $500-$2000 range, computed {computed_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not compute count from data: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
