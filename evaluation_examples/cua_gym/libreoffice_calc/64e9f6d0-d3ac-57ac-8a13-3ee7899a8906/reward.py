"""
FINAL REWARD SCRIPT - SUCCESS
Task: I need the total revenue by sales channel aggregated in a new sheet named "Channel_Summary" with channels as column headers using Pivot Table functionality.
Generated: 2025-11-24 07:48:10
Status: success
Model: o3
Total Steps: 9
"""

import os
import openpyxl


def verify_task(file_path: str) -> float:
    """Verify that the workbook contains a pivot-style summary sheet with
    revenue totals by channel.

    Scoring (progressive, max 1.0):
        0.20 – sheet named 'Channel_Summary' exists
        0.20 – all channels from SalesData appear as column headers
        0.60 – all channel totals in the summary match the computed totals
    """
    print(f"Verifying workbook: {file_path}")
    max_score = 1.0
    score = 0.0

    # ------------------------------------------------------------------
    # 1. Load workbook --------------------------------------------------
    # ------------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"✗ Could not load workbook: {e}")
        return 0.0  # nothing else can be verified

    # ------------------------------------------------------------------
    # 2. Check presence of summary sheet --------------------------------
    # ------------------------------------------------------------------
    summary_name = "Channel_Summary"
    if summary_name in wb.sheetnames:
        ws_summary = wb[summary_name]
        score += 0.20
        print("✓ 'Channel_Summary' sheet found (0.20)")
    else:
        print("✗ 'Channel_Summary' sheet NOT found – stopping checks")
        return score

    # ------------------------------------------------------------------
    # 3. Compute expected totals from SalesData --------------------------
    # ------------------------------------------------------------------
    sales_name = "SalesData"
    if sales_name not in wb.sheetnames:
        print("✗ 'SalesData' sheet missing – cannot compute expected totals")
        return score

    ws_sales = wb[sales_name]
    headers = [c.value for c in ws_sales[1]]
    try:
        channel_idx = next(i for i, h in enumerate(headers) if str(h).strip().lower() == "channel")
        revenue_idx = next(i for i, h in enumerate(headers) if str(h).strip().lower() == "revenue")
    except StopIteration:
        print("✗ Could not locate 'Channel' or 'Revenue' columns in 'SalesData'")
        return score

    expected_totals = {}
    for row in ws_sales.iter_rows(min_row=2, values_only=True):
        channel = row[channel_idx]
        revenue = row[revenue_idx]
        if channel is None or revenue is None:
            continue
        try:
            revenue = float(revenue)
        except (TypeError, ValueError):
            continue
        expected_totals[channel] = expected_totals.get(channel, 0.0) + revenue

    print("Computed expected totals from 'SalesData':")
    for ch, val in expected_totals.items():
        print(f"  • {ch}: {val}")

    # ------------------------------------------------------------------
    # 4. Verify headers in summary sheet --------------------------------
    # ------------------------------------------------------------------
    summary_headers = [c.value for c in ws_summary[1] if c.value not in (None, "")]
    print("Headers in 'Channel_Summary':", summary_headers)

    missing_headers = [ch for ch in expected_totals if ch not in summary_headers]
    if not missing_headers:
        score += 0.20
        print("✓ All channels present as headers (0.20)")
    else:
        print("✗ Missing headers for channels:", missing_headers)

    # Create mapping header ➜ column index in the summary sheet
    header_to_col = {
        str(cell.value).strip(): idx
        for idx, cell in enumerate(ws_summary[1], start=1)
        if cell.value not in (None, "")
    }

    # ------------------------------------------------------------------
    # 5. Locate the Grand Total row -------------------------------------
    # ------------------------------------------------------------------
    total_row_idx = None
    for r in ws_summary.iter_rows(min_row=2):
        label = r[0].value
        if label and str(label).strip().lower().startswith("grand total"):
            total_row_idx = r[0].row
            break

    if total_row_idx is None:
        print("✗ No 'Grand Total' row found – cannot check totals")
        return score

    # ------------------------------------------------------------------
    # 6. Compare totals --------------------------------------------------
    # ------------------------------------------------------------------
    correct = 0
    for ch, expected in expected_totals.items():
        col = header_to_col.get(ch)
        if col is None:
            continue  # header missing already penalised above
        cell_val = ws_summary.cell(row=total_row_idx, column=col).value
        try:
            actual = float(cell_val)
        except (TypeError, ValueError):
            actual = None

        if actual is not None and abs(actual - expected) <= 0.01:
            print(f"✓ Total for {ch} correct ({actual})")
            correct += 1
        else:
            print(f"✗ Total for {ch} incorrect – expected {expected}, found {cell_val}")

    # proportional points for totals (max 0.60)
    totals_score = 0.60 * correct / max(len(expected_totals), 1)
    score += totals_score
    if correct == len(expected_totals):
        print(f"✓ All channel totals correct (0.60)")
    else:
        print(f"▶ {correct}/{len(expected_totals)} totals correct ({totals_score:.2f})")

    final_score = min(score, max_score)
    print(f"Final SCORE: {final_score}")
    return final_score


# ----------------------------------------------------------------------
# Auto-detect a target workbook in /home/user ---------------------------
# ----------------------------------------------------------------------

def find_workbook(root: str = "/home/user") -> str | None:
    for dirpath, _, filenames in os.walk(root):
        for f in filenames:
            if f.lower().endswith(".xlsx"):
                return os.path.join(dirpath, f)
    return None


if __name__ == "__main__":
    target = find_workbook()
    if not target:
        print("No .xlsx file found – reward 0.0")
        print("REWARD: 0.0")
    else:
        reward_value = verify_task(target)
        print(f"REWARD: {reward_value}")
