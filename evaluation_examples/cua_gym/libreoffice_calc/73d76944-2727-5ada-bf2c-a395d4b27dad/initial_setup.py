"""
Initial Setup: Patient data spreadsheet with ward summary and weekly admissions tables
Task ID: osworld_calc_dual_chart_separate_tables_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_dual_chart_separate_tables_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Main Sheet: Patient Data ---
    ws = wb.active
    ws.title = 'Patient Data'

    # --- Table 1: Ward Summary ---
    # Row 1: Column headers (Ward in A, Treatment Count in B)
    ws['A1'] = 'Ward'
    ws['B1'] = 'Treatment Count'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # Rows 2-8: Ward data (7 wards) — matching context: "rows 2-8"
    ward_data = [
        ('General Medicine', 312),
        ('Cardiology',       287),
        ('Orthopedics',      245),
        ('Oncology',         198),
        ('Pediatrics',       334),
        ('Neurology',        221),
        ('Emergency',        421),
    ]
    for i, (ward, count) in enumerate(ward_data, start=2):
        ws.cell(row=i, column=1, value=ward)
        ws.cell(row=i, column=2, value=count)

    # Rows 9 (empty gap)

    # --- Table 2: Weekly Patient Admissions ---
    # Row 10: Column headers (Week in D, Admissions in E)
    ws['D10'] = 'Week'
    ws['E10'] = 'Admissions'
    ws['D10'].font = Font(bold=True)
    ws['E10'].font = Font(bold=True)

    # Rows 11-24: Weekly admissions data (14 weeks) — matching context: "rows 11-24"
    weekly_data = [
        ('Week 1',  134),
        ('Week 2',  147),
        ('Week 3',  159),
        ('Week 4',  142),
        ('Week 5',  168),
        ('Week 6',  175),
        ('Week 7',  183),
        ('Week 8',  162),
        ('Week 9',  191),
        ('Week 10', 178),
        ('Week 11', 195),
        ('Week 12', 204),
        ('Week 13', 212),
        ('Week 14', 225),
    ]
    for i, (week, admissions) in enumerate(weekly_data, start=11):
        ws.cell(row=i, column=4, value=week)
        ws.cell(row=i, column=5, value=admissions)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    # NO charts in initial state — agent must create them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()

