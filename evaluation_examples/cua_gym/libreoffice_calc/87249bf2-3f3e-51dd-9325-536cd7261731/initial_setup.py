"""
Initial Setup: Create student_ids.xlsx and raw_scores.ods on the Desktop; open terminal.
Task ID: osworld_multi_apps_terminal_calc_012
Domain: libreoffice_calc (multi-app: terminal + calc)

Initial state:
- /home/user/Desktop/student_ids.xlsx  — one column: StudentID (15 rows)
- /home/user/Desktop/raw_scores.ods    — one column: RawScore  (15 rows)
- Terminal window open on desktop
"""

import os
import shlex
import subprocess
import time

import openpyxl

DESKTOP = '/home/user/Desktop'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_student_ids_xlsx():
    """Create student_ids.xlsx on Desktop with one column: StudentID."""
    output = os.path.join(DESKTOP, 'student_ids.xlsx')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'StudentIDs'

    # Header
    ws.cell(row=1, column=1, value='StudentID')

    # 15 realistic student IDs
    student_ids = [
        'S001', 'S002', 'S003', 'S004', 'S005',
        'S006', 'S007', 'S008', 'S009', 'S010',
        'S011', 'S012', 'S013', 'S014', 'S015',
    ]
    for r, sid in enumerate(student_ids, 2):
        ws.cell(row=r, column=1, value=sid)

    wb.save(output)
    print(f'Created: {output}')


def create_raw_scores_ods():
    """Create raw_scores.ods on Desktop with one column: RawScore.

    We use openpyxl to write an xlsx first, then convert to ods via LibreOffice
    headless. This ensures a proper ODS file that LibreOffice Calc can open.
    """
    tmp_xlsx = os.path.join(DESKTOP, 'raw_scores_tmp.xlsx')
    output_ods = os.path.join(DESKTOP, 'raw_scores.ods')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'RawScores'

    # Header
    ws.cell(row=1, column=1, value='RawScore')

    # 15 realistic raw scores (integers between 0-100, varied)
    raw_scores = [78, 92, 65, 88, 45, 73, 95, 61, 84, 57, 91, 70, 82, 48, 77]
    for r, score in enumerate(raw_scores, 2):
        ws.cell(row=r, column=1, value=score)

    wb.save(tmp_xlsx)

    # Convert to ODS via LibreOffice headless
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        [
            'libreoffice', '--headless', '--convert-to', 'ods',
            '--outdir', DESKTOP, tmp_xlsx,
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
        timeout=60,
    )
    print('LibreOffice convert stdout:', result.stdout.decode())
    print('LibreOffice convert stderr:', result.stderr.decode())

    # The converted file will be named raw_scores_tmp.ods; rename to raw_scores.ods
    converted = os.path.join(DESKTOP, 'raw_scores_tmp.ods')
    if os.path.exists(converted):
        if os.path.exists(output_ods):
            os.remove(output_ods)
        os.rename(converted, output_ods)
        print(f'Created: {output_ods}')
    else:
        print(f'WARNING: ODS conversion may have failed. Expected: {converted}')
        # Fallback: try to list what was produced
        import glob
        files = glob.glob(os.path.join(DESKTOP, '*.ods'))
        print(f'ODS files in Desktop: {files}')

    # Remove the temporary xlsx
    if os.path.exists(tmp_xlsx):
        os.remove(tmp_xlsx)
        print(f'Removed temp file: {tmp_xlsx}')


def main():
    os.makedirs(DESKTOP, exist_ok=True)

    # Create the two source files
    create_student_ids_xlsx()
    create_raw_scores_ods()

    # Open a terminal (gnome-terminal) so the agent can start working
    launch_gui('gnome-terminal', delay_sec=2.0)
    print('GUI_READY: launched gnome-terminal with DISPLAY=:0')


main()
