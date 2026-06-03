"""
Initial Setup: CV Professors spreadsheet with Email, Lab/Group, Latest_Publication_Year columns empty
Task ID: osworld_multi_apps_web_prof_email_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_008'
OUTPUT = f'{WORKDIR}/CV_Professors.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Professors"

    # --- Column Headers ---
    headers = ['Name', 'Homepage', 'Email', 'Lab/Group', 'Latest_Publication_Year']
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # --- Professor Data ---
    # Name, Homepage, Email (empty), Lab/Group (empty), Latest_Publication_Year (empty)
    professors = [
        [
            "Fei-Fei Li",
            "https://profiles.stanford.edu/fei-fei-li",
            "",
            "",
            ""
        ],
        [
            "Trevor Darrell",
            "https://people.eecs.berkeley.edu/~trevor/",
            "",
            "",
            ""
        ],
        [
            "Deva Ramanan",
            "https://www.cs.cmu.edu/~deva/",
            "",
            "",
            ""
        ],
        [
            "Jitendra Malik",
            "https://people.eecs.berkeley.edu/~malik/",
            "",
            "",
            ""
        ],
        [
            "Antonio Torralba",
            "https://web.mit.edu/torralba/www/",
            "",
            "",
            ""
        ],
        [
            "Kaiming He",
            "https://kaiminghe.github.io/",
            "",
            "",
            ""
        ],
    ]

    thin = Side(style="thin", color="CCCCCC")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    name_font = Font(name="Calibri", size=11)

    for r, row_data in enumerate(professors, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = name_font
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 22
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch Chrome first (task requires web search)
    launch_gui('google-chrome --new-window', delay_sec=2.0)

    # Launch LibreOffice Calc with the spreadsheet
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
