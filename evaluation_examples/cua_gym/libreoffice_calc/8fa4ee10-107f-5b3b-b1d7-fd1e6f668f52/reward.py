"""
Reward Script: Verify that CV_Professors.xlsx has Email, Lab/Group, and Latest_Publication_Year
filled in for all 6 computer vision professors.
Task ID: osworld_multi_apps_web_prof_email_008
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.40): Email column filled for all 6 professors
  - Component 2 (0.40): Lab/Group column filled for all 6 professors
  - Component 3 (0.20): Latest_Publication_Year column filled for all 6 professors with valid year values
Total: 1.0

Initial state: Email, Lab/Group, and Latest_Publication_Year columns are all empty (None).
Golden state: All three columns are filled for all 6 rows.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_web_prof_email_008'
FILE_PATH = f'{WORKDIR}/CV_Professors.xlsx'

# Expected number of data rows (6 professors)
EXPECTED_ROW_COUNT = 6

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task asks the agent to fill in Email, Lab/Group, and Latest_Publication_Year
    for 6 computer vision professors. Initial state has these 3 columns empty.
    We only score the task-introduced changes (filled cells).
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the active sheet (should be 'Professors')
    try:
        if 'Professors' in wb.sheetnames:
            ws = wb['Professors']
        else:
            ws = wb.active
        print(f"INFO: Using sheet '{ws.title}'")
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify header row to confirm file structure (precondition gate, not scored)
    try:
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        expected_headers = ['Name', 'Homepage', 'Email', 'Lab/Group', 'Latest_Publication_Year']
        if headers != expected_headers:
            print(f"CRITICAL: Unexpected headers: {headers}. Expected: {expected_headers}")
            print("REWARD: 0.0")
            return 0.0
        print(f"INFO: Headers verified: {headers}")
    except Exception as e:
        print(f"CRITICAL: Cannot read headers: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Collect data rows (rows 2-7 for 6 professors)
    try:
        data_rows = []
        for row_idx in range(2, 2 + EXPECTED_ROW_COUNT):
            row_data = {
                'name': ws.cell(row=row_idx, column=1).value,
                'homepage': ws.cell(row=row_idx, column=2).value,
                'email': ws.cell(row=row_idx, column=3).value,
                'lab_group': ws.cell(row=row_idx, column=4).value,
                'latest_pub_year': ws.cell(row=row_idx, column=5).value,
            }
            data_rows.append(row_data)
        print(f"INFO: Found {len(data_rows)} data rows")
    except Exception as e:
        print(f"CRITICAL: Cannot read data rows: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: All professor names must be present (gate)
    try:
        professor_names = [r['name'] for r in data_rows]
        if not all(name for name in professor_names):
            print(f"CRITICAL: Some professor names are missing: {professor_names}")
            print("REWARD: 0.0")
            return 0.0
        print(f"INFO: All professor names present: {professor_names}")
    except Exception as e:
        print(f"CRITICAL: Cannot validate professor names: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------
    # Component 1: Email column filled for all 6 professors (0.40 points)
    # In the initial state, ALL email cells are None.
    # In the golden state, ALL email cells contain valid email addresses.
    # -----------------------------------------------------------
    try:
        email_filled = []
        email_missing = []
        for r in data_rows:
            email = r['email']
            if email and isinstance(email, str) and '@' in email and len(email.strip()) > 3:
                email_filled.append((r['name'], email))
            else:
                email_missing.append((r['name'], email))

        if len(email_filled) == EXPECTED_ROW_COUNT:
            print(f"PASS: Component 1 — All {EXPECTED_ROW_COUNT} Email cells filled with valid addresses (0.40 pts)")
            for name, email in email_filled:
                print(f"  {name}: {email}")
            total_score += 0.40
        elif len(email_filled) > 0:
            # Partial credit: proportional to number of emails filled
            partial = round((len(email_filled) / EXPECTED_ROW_COUNT) * 0.40, 4)
            print(f"PARTIAL: Component 1 — {len(email_filled)}/{EXPECTED_ROW_COUNT} Email cells filled (+{partial} pts)")
            for name, email in email_filled:
                print(f"  FILLED: {name}: {email}")
            for name, val in email_missing:
                print(f"  MISSING: {name}: {val}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No Email cells filled (0.00 pts)")
            for name, val in email_missing:
                print(f"  MISSING: {name}: {val}")
    except Exception as e:
        print(f"ERROR: Component 1 (Email) — {e}")

    # -----------------------------------------------------------
    # Component 2: Lab/Group column filled for all 6 professors (0.40 points)
    # In the initial state, ALL Lab/Group cells are None.
    # In the golden state, ALL Lab/Group cells contain non-empty lab/group names.
    # -----------------------------------------------------------
    try:
        lab_filled = []
        lab_missing = []
        for r in data_rows:
            lab = r['lab_group']
            if lab and isinstance(lab, str) and len(lab.strip()) > 2:
                lab_filled.append((r['name'], lab))
            else:
                lab_missing.append((r['name'], lab))

        if len(lab_filled) == EXPECTED_ROW_COUNT:
            print(f"PASS: Component 2 — All {EXPECTED_ROW_COUNT} Lab/Group cells filled (0.40 pts)")
            for name, lab in lab_filled:
                print(f"  {name}: {lab}")
            total_score += 0.40
        elif len(lab_filled) > 0:
            # Partial credit: proportional to number of labs filled
            partial = round((len(lab_filled) / EXPECTED_ROW_COUNT) * 0.40, 4)
            print(f"PARTIAL: Component 2 — {len(lab_filled)}/{EXPECTED_ROW_COUNT} Lab/Group cells filled (+{partial} pts)")
            for name, lab in lab_filled:
                print(f"  FILLED: {name}: {lab}")
            for name, val in lab_missing:
                print(f"  MISSING: {name}: {val}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No Lab/Group cells filled (0.00 pts)")
            for name, val in lab_missing:
                print(f"  MISSING: {name}: {val}")
    except Exception as e:
        print(f"ERROR: Component 2 (Lab/Group) — {e}")

    # -----------------------------------------------------------
    # Component 3: Latest_Publication_Year column filled for all 6 professors (0.20 points)
    # In the initial state, ALL Latest_Publication_Year cells are None.
    # In the golden state, ALL Latest_Publication_Year cells contain valid year values (integers >= 2000).
    # -----------------------------------------------------------
    try:
        year_filled = []
        year_missing = []
        for r in data_rows:
            year_val = r['latest_pub_year']
            is_valid_year = False
            if year_val is not None:
                try:
                    year_int = int(year_val)
                    if 2000 <= year_int <= 2030:
                        is_valid_year = True
                except (ValueError, TypeError):
                    pass
            if is_valid_year:
                year_filled.append((r['name'], year_val))
            else:
                year_missing.append((r['name'], year_val))

        if len(year_filled) == EXPECTED_ROW_COUNT:
            print(f"PASS: Component 3 — All {EXPECTED_ROW_COUNT} Latest_Publication_Year cells filled with valid years (0.20 pts)")
            for name, year in year_filled:
                print(f"  {name}: {year}")
            total_score += 0.20
        elif len(year_filled) > 0:
            # Partial credit: proportional to number of years filled
            partial = round((len(year_filled) / EXPECTED_ROW_COUNT) * 0.20, 4)
            print(f"PARTIAL: Component 3 — {len(year_filled)}/{EXPECTED_ROW_COUNT} Latest_Publication_Year cells filled (+{partial} pts)")
            for name, year in year_filled:
                print(f"  FILLED: {name}: {year}")
            for name, val in year_missing:
                print(f"  MISSING: {name}: {val}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No Latest_Publication_Year cells filled (0.00 pts)")
            for name, val in year_missing:
                print(f"  MISSING: {name}: {val}")
    except Exception as e:
        print(f"ERROR: Component 3 (Latest_Publication_Year) — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
