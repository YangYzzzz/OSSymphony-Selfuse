"""
Initial Setup: Grade band named ranges task
Task ID: calc_edu_named_range_gradeband_028
Domain: libreoffice_calc

Creates a gradebook spreadsheet with:
  - Sheet 'Settings': grade boundary labels and values
  - Sheet 'Grades': 30 students with hardcoded IF grade formulas (no named ranges)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_named_range_gradeband_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Settings ---
    ws_settings = wb.active
    ws_settings.title = 'Settings'

    # Grade boundary labels and values
    grade_bands = [
        ('A_MIN', 90),
        ('B_MIN', 80),
        ('C_MIN', 70),
        ('D_MIN', 60),
        ('FAIL_MAX', 59),
    ]
    for row_idx, (label, value) in enumerate(grade_bands, 1):
        ws_settings.cell(row=row_idx, column=1, value=label)
        ws_settings.cell(row=row_idx, column=2, value=value)

    # --- Sheet 2: Grades ---
    ws_grades = wb.create_sheet('Grades')

    # Headers
    ws_grades.cell(row=1, column=1, value='Student')
    ws_grades.cell(row=1, column=2, value='Score')
    ws_grades.cell(row=1, column=3, value='Letter')

    # 30 realistic student names and scores
    students = [
        ('Aiden Foster', 95),
        ('Brianna Nguyen', 82),
        ('Carlos Rivera', 74),
        ('Diana Patel', 91),
        ('Ethan Kim', 63),
        ('Fiona Walsh', 88),
        ('George Thompson', 57),
        ('Hannah Lee', 77),
        ('Ivan Petrov', 85),
        ('Julia Santos', 92),
        ('Kevin Brooks', 69),
        ('Laura Maddox', 98),
        ('Michael Chen', 71),
        ('Natalie Woods', 84),
        ('Oscar Martinez', 55),
        ('Priya Sharma', 89),
        ('Quinn Zhang', 66),
        ('Rachel Evans', 93),
        ('Samuel Okafor', 78),
        ('Tara Mitchell', 81),
        ('Ursula Grant', 62),
        ('Victor Huang', 96),
        ('Wendy Collins', 73),
        ('Xander Bell', 87),
        ('Yasmin Adeyemi', 58),
        ('Zachary Price', 79),
        ('Amara Diallo', 94),
        ('Brandon Hayes', 67),
        ('Chloe Robinson', 83),
        ('Derek Owens', 76),
    ]

    for row_idx, (name, score) in enumerate(students, 2):
        ws_grades.cell(row=row_idx, column=1, value=name)
        ws_grades.cell(row=row_idx, column=2, value=score)
        # Hardcoded IF formula — no named ranges
        formula = (
            f'=IF(B{row_idx}>=90,"A",'
            f'IF(B{row_idx}>=80,"B",'
            f'IF(B{row_idx}>=70,"C",'
            f'IF(B{row_idx}>=60,"D","F"))))'
        )
        ws_grades.cell(row=row_idx, column=3, value=formula)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
