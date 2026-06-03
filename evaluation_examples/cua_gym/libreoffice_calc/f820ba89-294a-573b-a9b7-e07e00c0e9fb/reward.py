"""
Reward Script: Set up named ranges for grade boundary values and update grade formulas
Task ID: calc_edu_named_range_gradeband_028
Domain: libreoffice_calc
Scoring:
  Component 1: Named ranges A_MIN, B_MIN, C_MIN, D_MIN are defined (0.4 pts - 0.1 each)
  Component 2: Named ranges reference correct cells in Settings sheet (0.2 pts)
  Component 3: All 30 grade formula rows in Grades!C use named ranges not hardcoded numbers (0.4 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_named_range_gradeband_028'

# Required named ranges and their expected references
REQUIRED_NAMES = {
    'A_MIN': 'Settings!$B$1',
    'B_MIN': 'Settings!$B$2',
    'C_MIN': 'Settings!$B$3',
    'D_MIN': 'Settings!$B$4',
}

# All 4 grade boundary names that must appear in column C formulas
GRADE_NAMES = ['A_MIN', 'B_MIN', 'C_MIN', 'D_MIN']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both required sheets must exist
    if 'Grades' not in wb.sheetnames or 'Settings' not in wb.sheetnames:
        print("CRITICAL: Required sheets 'Grades' and/or 'Settings' are missing")
        print("REWARD: 0.0")
        return 0.0

    # ------------------------------------------------------------------
    # Component 1: Named ranges A_MIN, B_MIN, C_MIN, D_MIN are defined
    # (0.1 points each, max 0.4 points)
    # This FAILS on initial (no defined names) and PASSES on golden.
    # ------------------------------------------------------------------
    defined_names = wb.defined_names  # DefinedNameDict
    existing_keys = list(defined_names.keys())

    component1_score = 0.0
    for name in REQUIRED_NAMES:
        try:
            if name in existing_keys:
                print(f"PASS: Named range '{name}' is defined")
                component1_score += 0.1
            else:
                print(f"FAIL: Named range '{name}' is NOT defined (found: {existing_keys})")
        except Exception as e:
            print(f"ERROR: Checking named range '{name}': {e}")

    if component1_score > 0:
        total_score += component1_score
    print(f"Component 1 subtotal: {component1_score}/0.4")

    # ------------------------------------------------------------------
    # Component 2: Named ranges reference correct Settings sheet cells
    # (0.05 points each for A_MIN..D_MIN = 0.2 points total)
    # This FAILS on initial (names don't exist at all) and PASSES on golden.
    # ------------------------------------------------------------------
    component2_score = 0.0
    for name, expected_ref in REQUIRED_NAMES.items():
        try:
            if name not in existing_keys:
                print(f"FAIL: '{name}' not defined, cannot check reference")
                continue
            dn = defined_names[name]
            actual_ref = dn.attr_text if hasattr(dn, 'attr_text') else str(dn)
            # Normalize comparison: case-insensitive, strip whitespace
            actual_norm = actual_ref.strip().upper()
            expected_norm = expected_ref.strip().upper()
            if actual_norm == expected_norm:
                print(f"PASS: '{name}' references correct cell: {actual_ref}")
                component2_score += 0.05
            else:
                print(f"FAIL: '{name}' expected ref={expected_ref}, actual={actual_ref}")
        except Exception as e:
            print(f"ERROR: Checking reference for '{name}': {e}")

    if component2_score > 0:
        total_score += component2_score
    print(f"Component 2 subtotal: {component2_score}/0.2")

    # ------------------------------------------------------------------
    # Component 3: All 30 grade formula rows (C2:C31) in Grades sheet
    # use named ranges instead of hardcoded numbers.
    # Award 0.4 pts only when ALL 30 rows are updated (no partial within this).
    # This FAILS on initial (all 30 rows have hardcoded numbers like 90, 80...)
    # and PASSES on golden (all 30 rows reference A_MIN, B_MIN, C_MIN, D_MIN).
    # ------------------------------------------------------------------
    component3_score = 0.0
    try:
        ws_grades = wb['Grades']
        named_rows = 0
        hardcoded_rows = 0
        for row in range(2, 32):  # rows 2 through 31 (30 students)
            cell_val = ws_grades.cell(row=row, column=3).value
            if cell_val is None:
                hardcoded_rows += 1
                print(f"FAIL: C{row} is empty, expected updated formula")
                continue
            formula_str = str(cell_val)
            # Check that ALL four grade boundary names appear in the formula
            uses_all_names = all(gn in formula_str for gn in GRADE_NAMES)
            # Check that hardcoded numbers 90, 80, 70, 60 are NOT used
            # (as comparison thresholds — they should not appear as numeric literals)
            has_hardcoded = any(f'>={n}' in formula_str or f'>= {n}' in formula_str
                                for n in ['90', '80', '70', '60'])
            if uses_all_names and not has_hardcoded:
                named_rows += 1
            else:
                hardcoded_rows += 1
                if hardcoded_rows <= 3:  # limit verbose output
                    print(f"FAIL: C{row} still uses hardcoded numbers or missing names: {formula_str[:80]}")

        if named_rows == 30:
            print(f"PASS: All 30 formula rows use named ranges (A_MIN, B_MIN, C_MIN, D_MIN)")
            component3_score = 0.4
        elif named_rows > 0:
            # Partial but we designed this as all-or-nothing at 0.4;
            # however, award proportional if most rows are updated
            fraction = named_rows / 30.0
            partial = round(fraction * 0.4, 2)
            print(f"PARTIAL: {named_rows}/30 formula rows use named ranges (score: {partial})")
            component3_score = partial
        else:
            print(f"FAIL: 0/30 formula rows use named ranges")
    except Exception as e:
        print(f"ERROR: Component 3 - checking formula column: {e}")

    if component3_score > 0:
        total_score += component3_score
    print(f"Component 3 subtotal: {component3_score}/0.4")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
