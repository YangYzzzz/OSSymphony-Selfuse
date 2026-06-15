"""
Initial Setup: Budget vs Actual spreadsheet with conditional formatting task
Task ID: calc_gcv_022
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget_vs_Actual"

    # Headers
    headers = ["Department", "Budget Amount", "Actual Spending"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # 44 departments with realistic data
    # Some departments overspend by >20%, others are under or close to budget
    departments = [
        ("Human Resources", 85000, 78500),
        ("Engineering", 195000, 245000),        # >20% over
        ("Marketing", 120000, 115800),
        ("Sales", 150000, 188000),              # >20% over
        ("Finance", 95000, 91200),
        ("Legal", 78000, 97500),                # >20% over
        ("Operations", 110000, 108000),
        ("Customer Support", 65000, 82000),     # >20% over
        ("Research & Development", 200000, 198500),
        ("Product Management", 88000, 112000),  # >20% over
        ("Quality Assurance", 55000, 53200),
        ("Data Analytics", 72000, 91000),       # >20% over
        ("IT Infrastructure", 130000, 127500),
        ("Security", 98000, 124000),            # >20% over
        ("Design", 60000, 58700),
        ("Content Strategy", 45000, 57500),     # >20% over
        ("Public Relations", 52000, 50800),
        ("Supply Chain", 140000, 175000),       # >20% over
        ("Procurement", 38000, 36500),
        ("Training & Development", 42000, 54600),  # >20% over
        ("Facilities Management", 92000, 89300),
        ("Internal Audit", 48000, 46200),
        ("Compliance", 67000, 85000),           # >20% over
        ("Business Intelligence", 76000, 74100),
        ("Cloud Services", 155000, 196000),     # >20% over
        ("Mobile Development", 115000, 112800),
        ("DevOps", 105000, 133500),             # >20% over
        ("Technical Writing", 35000, 33900),
        ("UX Research", 58000, 73500),          # >20% over
        ("Partnerships", 82000, 80100),
        ("Event Management", 28000, 36000),     # >20% over
        ("Corporate Strategy", 90000, 87600),
        ("Talent Acquisition", 54000, 69000),   # >20% over
        ("Benefits Administration", 32000, 31100),
        ("Risk Management", 71000, 90500),      # >20% over
        ("Investor Relations", 46000, 44800),
        ("Environmental Health", 39000, 50000), # >20% over
        ("Logistics", 125000, 122000),
        ("Fleet Management", 63000, 80000),     # >20% over
        ("Payroll", 41000, 39800),
        ("Real Estate", 170000, 168500),
        ("Mergers & Acquisitions", 180000, 228000),  # >20% over
        ("Government Relations", 56000, 54500),
        ("International Ops", 145000, 184000),  # >20% over
    ]

    for r, (dept, budget, actual) in enumerate(departments, 2):
        ws.cell(row=r, column=1, value=dept).border = thin_border
        budget_cell = ws.cell(row=r, column=2, value=budget)
        budget_cell.number_format = '$#,##0'
        budget_cell.border = thin_border
        actual_cell = ws.cell(row=r, column=3, value=actual)
        actual_cell.number_format = '$#,##0'
        actual_cell.border = thin_border

    # Set column widths for readability
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
