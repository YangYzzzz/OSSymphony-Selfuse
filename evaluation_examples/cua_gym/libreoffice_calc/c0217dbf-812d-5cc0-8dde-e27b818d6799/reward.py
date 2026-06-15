"""
Reward Script: Create pivot table in Sheet2 showing grant totals by funding agency,
               with a styled merged header (blue fill, bold white text, 'Grant Funding Summary 2024').
Task ID: osworld_calc_pivot_multi_styled_007
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: Merged header cell A1:B1 with 'Grant Funding Summary 2024' text  — 0.35 pts
  Component 2: Header cell styling — blue fill (FF4472C4) + bold white font       — 0.25 pts
  Component 3: Pivot table column headers present in Sheet2                        — 0.20 pts
  Component 4: Pivot data contains correct agency totals                           — 0.20 pts
  Total: 1.0
"""

import os
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_007'

# Expected pivot values: Funding Agency -> Total Grant Amount
EXPECTED_PIVOT = {
    'DARPA': 3120000,
    'Department of Energy': 1730000,
    'National Aeronautics and Space Administration': 860000,
    'National Institutes of Health': 2800000,
    'National Science Foundation': 1460000,
}

# Required tolerance for numeric comparison
TOLERANCE = 0.01

# Blue fill color (ARGB) as set by setup-gen
EXPECTED_FILL_COLOR = 'FF4472C4'
# White font color (ARGB)
EXPECTED_FONT_COLOR = 'FFFFFFFF'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet2 must exist
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 does not exist in the workbook")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Sheet2']

    # -------------------------------------------------------------------------
    # Component 1: Merged header cell A1:B1 with correct text (0.35 points)
    # This MUST fail on initial_env (Sheet2 is empty) and pass on golden_env.
    # -------------------------------------------------------------------------
    try:
        # Check that A1:B1 is merged
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        is_merged_a1_b1 = 'A1:B1' in merged_ranges

        # Check B1 is a MergedCell (indicating A1:B1 merge)
        b1_is_merged = isinstance(ws['B1'], MergedCell)

        # Check the text in A1 (the top-left of the merge)
        header_text = ws['A1'].value
        header_correct = (
            header_text is not None
            and str(header_text).strip() == 'Grant Funding Summary 2024'
        )

        if is_merged_a1_b1 and b1_is_merged and header_correct:
            print(f"PASS: Component 1 — A1:B1 merged with text 'Grant Funding Summary 2024' (0.35 pts)")
            total_score += 0.35
        else:
            if not (is_merged_a1_b1 or b1_is_merged):
                print(f"FAIL: Component 1 — A1:B1 is not merged. Merged ranges: {merged_ranges}")
            elif not header_correct:
                print(f"FAIL: Component 1 — A1 text expected 'Grant Funding Summary 2024', found: {repr(header_text)}")
            else:
                print(f"FAIL: Component 1 — merge or header text not correct")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Header cell A1 styling — blue fill + bold white font (0.25 points)
    # This MUST fail on initial_env (Sheet2 A1 is empty/unstyled).
    # -------------------------------------------------------------------------
    try:
        cell_a1 = ws['A1']

        # Check bold
        is_bold = cell_a1.font.bold is True

        # Check white font color (FFFFFFFF)
        font_color_ok = False
        try:
            font_rgb = cell_a1.font.color.rgb
            # Accept FFFFFFFF (white with full alpha) or 00FFFFFF (white with zero alpha)
            font_color_ok = (
                font_rgb is not None
                and font_rgb.upper() in {'FFFFFFFF', '00FFFFFF'}
            )
        except Exception:
            font_color_ok = False

        # Check blue fill (FF4472C4 or similar blue)
        fill_color_ok = False
        try:
            fill_rgb = cell_a1.fill.fgColor.rgb
            # Accept exact match for FF4472C4
            fill_color_ok = fill_rgb is not None and fill_rgb.upper() == EXPECTED_FILL_COLOR.upper()
        except Exception:
            fill_color_ok = False

        if is_bold and font_color_ok and fill_color_ok:
            print(f"PASS: Component 2 — A1 has bold white font and blue fill (0.25 pts)")
            total_score += 0.25
        else:
            details = []
            if not is_bold:
                details.append(f"bold={cell_a1.font.bold}")
            if not font_color_ok:
                try:
                    details.append(f"font_color={cell_a1.font.color.rgb}")
                except Exception:
                    details.append("font_color=error")
            if not fill_color_ok:
                try:
                    details.append(f"fill_color={cell_a1.fill.fgColor.rgb}")
                except Exception:
                    details.append("fill_color=error")
            print(f"FAIL: Component 2 — Header styling not correct: {', '.join(details)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Pivot table column headers present (0.20 points)
    # Row 2 should have 'Funding Agency' in column A and 'Total Grant Amount' in column B.
    # This MUST fail on initial_env (Sheet2 has no such headers).
    # -------------------------------------------------------------------------
    try:
        header_row_a = ws['A2'].value
        header_row_b = ws['B2'].value

        agency_header_ok = (
            header_row_a is not None
            and 'funding agency' in str(header_row_a).strip().lower()
        )
        amount_header_ok = (
            header_row_b is not None
            and 'grant amount' in str(header_row_b).strip().lower()
        )

        if agency_header_ok and amount_header_ok:
            print(f"PASS: Component 3 — Pivot headers present: A2={repr(header_row_a)}, B2={repr(header_row_b)} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Expected 'Funding Agency' in A2 and 'Total Grant Amount' in B2. "
                  f"Found A2={repr(header_row_a)}, B2={repr(header_row_b)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Pivot data contains correct agency totals (0.20 points)
    # Rows 3 onwards should have each funding agency and its summed amount.
    # This MUST fail on initial_env (Sheet2 has no data).
    # -------------------------------------------------------------------------
    try:
        # Read all data rows from row 3 onwards
        actual_pivot = {}
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=2, values_only=True):
            agency, amount = row
            if agency is not None:
                actual_pivot[str(agency).strip()] = amount

        if not actual_pivot:
            print(f"FAIL: Component 4 — No pivot data rows found in Sheet2 (rows 3+)")
        else:
            matched = 0
            total_expected = len(EXPECTED_PIVOT)
            mismatches = []

            for agency, expected_total in EXPECTED_PIVOT.items():
                if agency in actual_pivot:
                    actual_val = actual_pivot[agency]
                    if actual_val is not None:
                        try:
                            if abs(float(actual_val) - float(expected_total)) <= TOLERANCE:
                                matched += 1
                            else:
                                mismatches.append(
                                    f"{agency}: expected {expected_total}, got {actual_val}"
                                )
                        except (TypeError, ValueError):
                            mismatches.append(f"{agency}: non-numeric value {repr(actual_val)}")
                    else:
                        mismatches.append(f"{agency}: None value")
                else:
                    mismatches.append(f"{agency}: not found in pivot")

            # Full credit if all agencies are correct
            if matched == total_expected:
                print(f"PASS: Component 4 — All {matched}/{total_expected} agency totals correct (0.20 pts)")
                total_score += 0.20
            else:
                # Partial credit: at least half correct (0.10 pts)
                half_threshold = total_expected // 2
                partial_ok = matched >= half_threshold
                if partial_ok:
                    print(f"PARTIAL: Component 4 — {matched}/{total_expected} agency totals correct (0.10 pts). "
                          f"Issues: {mismatches}")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 4 — Only {matched}/{total_expected} agency totals correct. "
                          f"Issues: {mismatches}")

    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against canonical artifact path on VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
