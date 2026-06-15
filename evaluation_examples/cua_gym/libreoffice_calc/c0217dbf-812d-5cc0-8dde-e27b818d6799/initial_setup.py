"""
Initial Setup: Research grant records spreadsheet (pre-task state)
Task ID: osworld_calc_pivot_multi_styled_007
Domain: libreoffice_calc

Creates Sheet1 with research grant data and an empty Sheet2.
The agent must create a styled header and pivot table in Sheet2.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_pivot_multi_styled_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Grant Records ---
    ws1 = wb.active
    ws1.title = "Sheet1"

    # Headers
    headers = ['Grant ID', 'Funding Agency', 'Department', 'Principal Investigator', 'Grant Amount', 'Year']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Realistic grant data
    data = [
        ['GR-2024-001', 'National Science Foundation', 'Biology', 'Dr. Elena Vasquez', 285000, 2024],
        ['GR-2024-002', 'National Institutes of Health', 'Medicine', 'Dr. James Okafor', 520000, 2024],
        ['GR-2024-003', 'Department of Energy', 'Physics', 'Dr. Yuki Tanaka', 175000, 2024],
        ['GR-2024-004', 'National Science Foundation', 'Chemistry', 'Dr. Priya Mehta', 310000, 2024],
        ['GR-2024-005', 'National Institutes of Health', 'Neuroscience', 'Dr. Carlos Rivera', 445000, 2024],
        ['GR-2024-006', 'DARPA', 'Computer Science', 'Dr. Sarah Kowalski', 890000, 2024],
        ['GR-2024-007', 'National Science Foundation', 'Environmental Science', 'Dr. Marcus Webb', 230000, 2024],
        ['GR-2024-008', 'Department of Energy', 'Materials Science', 'Dr. Aisha Patel', 395000, 2024],
        ['GR-2024-009', 'National Institutes of Health', 'Pharmacology', 'Dr. Thomas Nguyen', 610000, 2024],
        ['GR-2024-010', 'DARPA', 'Robotics', 'Dr. Fatima Al-Hassan', 1250000, 2024],
        ['GR-2024-011', 'National Science Foundation', 'Mathematics', 'Dr. Robert Chen', 195000, 2024],
        ['GR-2024-012', 'National Aeronautics and Space Administration', 'Astronomy', 'Dr. Lucía Fernandez', 475000, 2024],
        ['GR-2024-013', 'Department of Energy', 'Renewable Energy', 'Dr. Benjamin Adeyemi', 540000, 2024],
        ['GR-2024-014', 'National Institutes of Health', 'Oncology', 'Dr. Mia Johansson', 730000, 2024],
        ['GR-2024-015', 'National Science Foundation', 'Geology', 'Dr. Arjun Sharma', 165000, 2024],
        ['GR-2024-016', 'DARPA', 'Cybersecurity', 'Dr. Olivia Petrov', 980000, 2024],
        ['GR-2024-017', 'National Aeronautics and Space Administration', 'Planetary Science', 'Dr. Noah Kim', 385000, 2024],
        ['GR-2024-018', 'National Science Foundation', 'Cognitive Science', 'Dr. Isabella Cruz', 275000, 2024],
        ['GR-2024-019', 'Department of Energy', 'Nuclear Engineering', 'Dr. Ethan Larsson', 620000, 2024],
        ['GR-2024-020', 'National Institutes of Health', 'Immunology', 'Dr. Zara Ahmed', 495000, 2024],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=10)
            if c == 5:  # Grant Amount column — currency format
                cell.number_format = '$#,##0'

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 16
    ws1.column_dimensions['B'].width = 45
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 30
    ws1.column_dimensions['E'].width = 16
    ws1.column_dimensions['F'].width = 8

    # Freeze header row
    ws1.freeze_panes = "A2"

    # --- Sheet 2: Empty (agent will create pivot table here) ---
    ws2 = wb.create_sheet("Sheet2")
    # Leave Sheet2 completely empty — agent must add styled header and pivot table

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
