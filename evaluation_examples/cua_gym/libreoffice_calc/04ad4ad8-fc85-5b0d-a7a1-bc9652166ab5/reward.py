"""
Reward Script: Create pivot table from financial data with currency formatting
Task ID: calc_pivot_039
Domain: libreoffice_calc
Scoring:
  Component 1: Pivot table sheet exists (0.15)
  Component 2: Correct row labels / structure (0.25)
  Component 3: Correct SUM values per account (0.35)
  Component 4: Currency formatting on amount cells (0.25)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_039'

# Expected pivot table values (from task context ground truth)
EXPECTED_ACCOUNTS = {
    'Revenue': 245000,
    'COGS': 128500,
    'OpEx': 65200,
    'CapEx': 42300,
}
GRAND_TOTAL = 481000
CURRENCY_FORMAT = '$#,##0.00'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A pivot table sheet exists (separate from Finance) (0.15 points)
    # The initial file only has 'Finance'. The golden adds a pivot table sheet.
    try:
        pivot_sheet = None
        for sn in wb.sheetnames:
            if sn.lower() != 'finance':
                ws_candidate = wb[sn]
                # Check if it has at least a header row with account-like structure
                if ws_candidate.max_row >= 2 and ws_candidate.max_column >= 2:
                    pivot_sheet = ws_candidate
                    break
        if pivot_sheet is not None:
            print(f"PASS: Component 1 — Pivot table sheet found: '{pivot_sheet.title}' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — No pivot table sheet found (only sheets: {wb.sheetnames})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_sheet is None:
        # Cannot proceed without a pivot table sheet
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Component 2: Correct row labels — all 4 accounts present as row values (0.25 points)
    try:
        # Scan column A for account labels
        found_accounts = set()
        account_rows = {}  # account_name -> row number
        for r in range(1, pivot_sheet.max_row + 1):
            val = pivot_sheet.cell(row=r, column=1).value
            if val is not None:
                val_str = str(val).strip()
                if val_str in EXPECTED_ACCOUNTS:
                    found_accounts.add(val_str)
                    account_rows[val_str] = r

        matched = len(found_accounts)
        expected = len(EXPECTED_ACCOUNTS)
        if matched == expected:
            print(f"PASS: Component 2 — All 4 account labels found: {found_accounts} (0.25 pts)")
            total_score += 0.25
        elif matched > 0:
            partial = round(0.25 * matched / expected, 3)
            print(f"PARTIAL: Component 2 — {matched}/{expected} accounts found: {found_accounts} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No expected account labels found in pivot table column A")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct SUM values for each account (0.35 points)
    # Each account check is worth 0.35/4 = 0.0875 points
    try:
        points_per_account = 0.35 / len(EXPECTED_ACCOUNTS)
        comp3_score = 0.0

        for acct_name, expected_val in EXPECTED_ACCOUNTS.items():
            if acct_name not in account_rows:
                print(f"FAIL: Component 3 — Account '{acct_name}' row not found, cannot check value")
                continue

            row = account_rows[acct_name]
            # Find the numeric value in column B (or scan columns for the amount)
            found_val = None
            for c in range(2, pivot_sheet.max_column + 1):
                cell_val = pivot_sheet.cell(row=row, column=c).value
                if cell_val is not None:
                    try:
                        found_val = float(cell_val)
                        break
                    except (ValueError, TypeError):
                        continue

            if found_val is not None and abs(found_val - expected_val) < 1.0:
                print(f"PASS: Component 3 — {acct_name}: expected {expected_val}, found {found_val} ({points_per_account:.4f} pts)")
                comp3_score += points_per_account
            else:
                print(f"FAIL: Component 3 — {acct_name}: expected {expected_val}, found {found_val}")

        if comp3_score > 0:
            total_score += comp3_score
        print(f"  Component 3 subtotal: {comp3_score:.4f}/0.35")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Currency formatting ($#,##0.00) on amount cells (0.25 points)
    try:
        formatted_count = 0
        total_amount_cells = 0

        for acct_name in EXPECTED_ACCOUNTS:
            if acct_name not in account_rows:
                continue
            row = account_rows[acct_name]
            for c in range(2, pivot_sheet.max_column + 1):
                cell = pivot_sheet.cell(row=row, column=c)
                if cell.value is not None:
                    try:
                        float(cell.value)
                        total_amount_cells += 1
                        # Check for currency format — accept variants that include $ and decimals
                        nf = cell.number_format
                        if nf and '$' in nf and '0.00' in nf:
                            formatted_count += 1
                        else:
                            print(f"FAIL: Component 4 — Cell {cell.coordinate} format is '{nf}', expected currency with $")
                    except (ValueError, TypeError):
                        pass

        if total_amount_cells > 0 and formatted_count == total_amount_cells:
            print(f"PASS: Component 4 — All {formatted_count} amount cells have currency formatting (0.25 pts)")
            total_score += 0.25
        elif formatted_count > 0:
            partial = round(0.25 * formatted_count / max(total_amount_cells, 1), 3)
            print(f"PARTIAL: Component 4 — {formatted_count}/{total_amount_cells} cells formatted as currency ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — No amount cells have currency formatting")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
