"""
Initial Setup: Create financial transaction data for pivot table task
Task ID: calc_pivot_039
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Finance'

    # --- Headers ---
    headers = ['TxnID', 'Date', 'Account', 'SubAccount', 'Amount']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Define account structure ---
    # Target totals:
    #   Revenue  = 245000.00
    #   COGS     = 128500.00
    #   OpEx     =  65200.00
    #   CapEx    =  42300.00
    # Total transactions = 180

    account_config = {
        'Revenue': {
            'target': 245000.00,
            'count': 60,
            'subaccounts': [
                'Product Sales', 'Service Revenue', 'Licensing Fees',
                'Subscription Income', 'Consulting Revenue'
            ]
        },
        'COGS': {
            'target': 128500.00,
            'count': 50,
            'subaccounts': [
                'Raw Materials', 'Direct Labor', 'Manufacturing Overhead',
                'Shipping Costs', 'Packaging'
            ]
        },
        'OpEx': {
            'target': 65200.00,
            'count': 40,
            'subaccounts': [
                'Office Rent', 'Utilities', 'Marketing',
                'Travel & Entertainment', 'Software Licenses'
            ]
        },
        'CapEx': {
            'target': 42300.00,
            'count': 30,
            'subaccounts': [
                'Equipment Purchase', 'IT Infrastructure',
                'Facility Upgrades', 'Vehicle Acquisition'
            ]
        },
    }

    # Generate transactions
    start_date = datetime(2024, 1, 5)
    row_idx = 2
    txn_id = 1

    for account, cfg in account_config.items():
        target = cfg['target']
        count = cfg['count']
        subs = cfg['subaccounts']

        # Generate random amounts that sum to target
        # Create count-1 random proportions, then compute last to hit target exactly
        raw = [random.uniform(0.5, 2.0) for _ in range(count - 1)]
        total_raw = sum(raw)
        amounts = [round(r / total_raw * target, 2) for r in raw]
        last_amount = round(target - sum(amounts), 2)
        amounts.append(last_amount)
        random.shuffle(amounts)

        for i in range(count):
            date = start_date + timedelta(days=random.randint(0, 355))
            sub = random.choice(subs)
            amt = amounts[i]

            ws.cell(row=row_idx, column=1, value=txn_id)
            ws.cell(row=row_idx, column=2, value=date.strftime('%Y-%m-%d'))
            ws.cell(row=row_idx, column=3, value=account)
            ws.cell(row=row_idx, column=4, value=sub)
            ws.cell(row=row_idx, column=5, value=amt)

            txn_id += 1
            row_idx += 1

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Verify sums
    import openpyxl as ox
    wb2 = ox.load_workbook(OUTPUT)
    ws2 = wb2['Finance']
    sums = {}
    for r in range(2, 182):
        acct = ws2.cell(row=r, column=3).value
        amt = ws2.cell(row=r, column=5).value
        sums[acct] = sums.get(acct, 0) + amt
    for acct, total in sorted(sums.items()):
        print(f'  {acct}: {total:.2f}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
