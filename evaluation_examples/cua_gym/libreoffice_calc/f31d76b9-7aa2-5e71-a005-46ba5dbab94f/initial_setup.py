"""
Initial Setup: Project Tasks spreadsheet with no conditional formatting
Task ID: calc_fmt_conditional_formula_based_081
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_conditional_formula_based_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Project Tasks'

    # --- Headers ---
    headers = ['ID', 'Task Name', 'Start', 'End', 'Status']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16

    # --- Data rows 2-20: realistic project tasks ---
    data = [
        [1,  'Redesign company homepage',           '2025-01-05', '2025-01-20', 'Complete'],
        [2,  'Set up CI/CD pipeline',               '2025-01-08', '2025-02-01', 'Complete'],
        [3,  'Migrate database to PostgreSQL',      '2025-01-10', '2025-03-15', 'In Progress'],
        [4,  'Conduct Q1 user research interviews', '2025-01-15', '2025-02-10', 'Complete'],
        [5,  'Draft product roadmap for 2025',      '2025-01-20', '2025-02-05', 'Complete'],
        [6,  'Implement OAuth2 authentication',     '2025-01-22', '2025-03-01', 'In Progress'],
        [7,  'Write API documentation',             '2025-02-01', '2025-03-20', 'Not Started'],
        [8,  'Performance optimization audit',      '2025-02-03', '2025-02-28', 'Complete'],
        [9,  'Onboard new marketing team members',  '2025-02-05', '2025-02-15', 'Complete'],
        [10, 'Build analytics dashboard',           '2025-02-10', '2025-04-01', 'In Progress'],
        [11, 'Update privacy policy documents',     '2025-02-12', '2025-02-20', 'Complete'],
        [12, 'Deploy mobile app v2.3',              '2025-02-15', '2025-04-10', 'Not Started'],
        [13, 'Conduct accessibility review',        '2025-02-18', '2025-03-05', 'In Progress'],
        [14, 'Integrate Stripe payment gateway',    '2025-02-20', '2025-03-25', 'Not Started'],
        [15, 'Automate weekly reporting',           '2025-02-22', '2025-03-10', 'Complete'],
        [16, 'Set up monitoring and alerting',      '2025-02-25', '2025-03-30', 'In Progress'],
        [17, 'Archive 2024 project files',          '2025-03-01', '2025-03-08', 'Complete'],
        [18, 'Plan Q2 sprint schedule',             '2025-03-03', '2025-03-12', 'Not Started'],
        [19, 'Refactor authentication module',      '2025-03-05', '2025-04-15', 'Not Started'],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        ws.cell(row=r, column=3, value=row_data[2])
        ws.cell(row=r, column=4, value=row_data[3])
        ws.cell(row=r, column=5, value=row_data[4])

    # NO conditional formatting in initial file

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Project Tasks')
    print(f'Rows: 1 header + 19 data rows (rows 2-20)')
    print(f'Status values: Complete, In Progress, Not Started')
    print(f'NO conditional formatting applied')


create_initial()
