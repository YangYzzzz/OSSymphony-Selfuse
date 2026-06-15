"""
Initial Setup: Build a KPI summary section in the dashboard sheet
Task ID: calc_gsd_010
Domain: libreoffice_calc

Creates an empty Summary sheet (where the agent will build the KPI section)
plus a Data sheet with raw quarterly figures for context.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary (empty - agent will populate this) ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    # Leave completely empty per task context

    # --- Sheet 2: Data (raw quarterly data for context/realism) ---
    ws_data = wb.create_sheet('Data')

    # Headers
    headers = ['Department', 'Q1 Revenue', 'Q2 Revenue', 'Q3 Revenue', 'Headcount', 'Satisfaction']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    thin_border = Border(bottom=Side(style="thin", color="000000"))

    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Realistic department data
    data = [
        ['Engineering',    1250000, 1340000, 1420000, 185, 91.2],
        ['Sales',           980000, 1050000, 1120000, 142, 84.7],
        ['Marketing',       420000,  445000,  470000,  68, 88.3],
        ['Operations',      310000,  325000,  340000,  95, 86.1],
        ['Customer Support', 180000,  195000,  210000, 112, 82.5],
        ['Human Resources',  150000,  155000,  160000,  38, 90.4],
        ['Finance',          220000,  230000,  240000,  45, 87.9],
        ['Legal',            130000,  135000,  140000,  22, 85.6],
        ['Product',          310000,  340000,  380000,  56, 89.1],
        ['Research',         280000,  295000,  320000,  34, 92.3],
        ['IT Infrastructure', 190000, 200000,  210000,  48, 83.8],
        ['Quality Assurance', 160000, 170000,  180000,  42, 87.2],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws_data.cell(row=r, column=c, value=val)

    # Format currency columns
    for r in range(2, len(data) + 2):
        for c in [2, 3, 4]:
            ws_data.cell(row=r, column=c).number_format = '$#,##0'

    # Format satisfaction column
    for r in range(2, len(data) + 2):
        ws_data.cell(row=r, column=6).number_format = '0.0'

    # Totals row
    total_row = len(data) + 2
    ws_data.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    for c in [2, 3, 4]:
        col_letter = chr(64 + c)
        ws_data.cell(row=total_row, column=c,
                     value=f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
        ws_data.cell(row=total_row, column=c).number_format = '$#,##0'
        ws_data.cell(row=total_row, column=c).font = Font(bold=True)
    ws_data.cell(row=total_row, column=5,
                 value=f'=SUM(E2:E{total_row - 1})').font = Font(bold=True)

    # Column widths for readability
    ws_data.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D']:
        ws_data.column_dimensions[col_letter].width = 15
    ws_data.column_dimensions['E'].width = 12
    ws_data.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
