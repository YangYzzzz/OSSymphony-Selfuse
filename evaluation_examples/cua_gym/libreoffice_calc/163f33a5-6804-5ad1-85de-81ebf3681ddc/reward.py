"""
Reward Script: Build a KPI summary section in rows 1-8 of the dashboard sheet
Task ID: calc_gsd_010
Domain: libreoffice_calc
Scoring:
  Component 1: Title in A1 merged A1:D1, 14pt bold (0.20)
  Component 2: KPI labels in A2:A8 (0.20)
  Component 3: KPI values in B2:B8 (0.20)
  Component 4: Currency formatting on B2:B5 (0.15)
  Component 5: Percentage formatting on B7:B8 (0.10)
  Component 6: Data bar conditional formatting on B2:B8 (0.15)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_010'

EXPECTED_LABELS = [
    'Total Revenue',
    'Gross Profit',
    'Operating Expenses',
    'Net Income',
    'Customer Count',
    'Satisfaction Score',
    'Market Share',
]

EXPECTED_VALUES = {
    'B2': 4250000,
    'B3': 1870000,
    'B4': 980000,
    'B5': 890000,
    'B6': 12450,
    'B7': 87.5,
    'B8': 23.4,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the Summary sheet
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0
    ws = wb['Summary']

    # Component 1: Title in A1 — merged A1:D1, text, 14pt bold (0.20 points)
    try:
        a1_val = ws['A1'].value
        has_title = (a1_val is not None and 'KPI Dashboard' in str(a1_val) and 'Q3 2024' in str(a1_val))

        # Check merge: B1 should be a MergedCell
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'D1' in r for r in merged_ranges)

        # Check font: bold and 14pt
        font = ws['A1'].font
        has_bold = (font.bold is True)
        has_14pt = (font.size is not None and abs(float(font.size) - 14.0) < 1.0)

        if has_title and has_merge and has_bold and has_14pt:
            print(f"PASS: Component 1 — Title '{a1_val}' merged A1:D1, bold={has_bold}, size={font.size} (0.20 pts)")
            total_score += 0.20
        else:
            details = f"title={has_title}, merge={has_merge}, bold={has_bold}, 14pt={has_14pt}"
            print(f"FAIL: Component 1 — {details}, A1 value={a1_val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: KPI labels in A2:A8 (0.20 points)
    try:
        matched_labels = 0
        for i, expected_label in enumerate(EXPECTED_LABELS):
            row = i + 2
            actual = ws.cell(row=row, column=1).value
            if actual is not None and str(actual).strip() == expected_label:
                matched_labels += 1
            else:
                print(f"  Label mismatch row {row}: expected '{expected_label}', got {actual!r}")

        if matched_labels == 7:
            print(f"PASS: Component 2 — All 7 KPI labels correct (0.20 pts)")
            total_score += 0.20
        elif matched_labels >= 4:
            partial = round(0.20 * matched_labels / 7, 2)
            print(f"PARTIAL: Component 2 — {matched_labels}/7 labels correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {matched_labels}/7 labels correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: KPI values in B2:B8 (0.20 points)
    try:
        matched_values = 0
        for coord, expected_val in EXPECTED_VALUES.items():
            actual = ws[coord].value
            if actual is not None:
                try:
                    if abs(float(actual) - expected_val) < 0.5:
                        matched_values += 1
                    else:
                        print(f"  Value mismatch {coord}: expected {expected_val}, got {actual}")
                except (ValueError, TypeError):
                    print(f"  Value type error {coord}: expected {expected_val}, got {actual!r}")
            else:
                print(f"  Value missing {coord}: expected {expected_val}")

        if matched_values == 7:
            print(f"PASS: Component 3 — All 7 KPI values correct (0.20 pts)")
            total_score += 0.20
        elif matched_values >= 4:
            partial = round(0.20 * matched_values / 7, 2)
            print(f"PARTIAL: Component 3 — {matched_values}/7 values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {matched_values}/7 values correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Currency formatting on B2:B5 — must contain '$' (0.15 points)
    try:
        currency_count = 0
        for coord in ['B2', 'B3', 'B4', 'B5']:
            nf = ws[coord].number_format
            if nf and '$' in str(nf):
                currency_count += 1
            else:
                print(f"  Currency format missing {coord}: got {nf!r}")

        if currency_count == 4:
            print(f"PASS: Component 4 — B2:B5 all have currency formatting (0.15 pts)")
            total_score += 0.15
        elif currency_count >= 2:
            partial = round(0.15 * currency_count / 4, 2)
            print(f"PARTIAL: Component 4 — {currency_count}/4 currency formatted ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {currency_count}/4 cells have currency format")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Percentage formatting on B7:B8 (0.10 points)
    try:
        pct_count = 0
        for coord in ['B7', 'B8']:
            nf = ws[coord].number_format
            if nf and '%' in str(nf):
                pct_count += 1
            else:
                print(f"  Percentage format missing {coord}: got {nf!r}")

        if pct_count == 2:
            print(f"PASS: Component 5 — B7:B8 both have percentage formatting (0.10 pts)")
            total_score += 0.10
        elif pct_count == 1:
            print(f"PARTIAL: Component 5 — 1/2 percentage formatted (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No percentage formatting found on B7:B8")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Data bar conditional formatting on B2:B8 (0.15 points)
    try:
        databar_range = None
        for cf in ws.conditional_formatting:
            cf_range_str = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    # Check that the range covers at least B2:B8
                    if 'B2' in cf_range_str or 'B2:B8' in cf_range_str or '$B$2' in cf_range_str:
                        databar_range = cf_range_str
                        print(f"  Found dataBar rule on range: {cf_range_str}")
                        break
            if databar_range is not None:
                break

        if databar_range is not None:
            print(f"PASS: Component 6 — Data bar conditional formatting found on B2:B8 (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 6 — No data bar conditional formatting found on B2:B8")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
