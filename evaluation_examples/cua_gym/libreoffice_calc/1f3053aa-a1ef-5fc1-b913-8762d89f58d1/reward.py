"""
Reward Script: Insert 'Assumptions' sheet from external file into workbook
Task ID: calc_gsi_087
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): 'Assumptions' sheet exists in the workbook
  Component 2 (0.20): Assumptions sheet has correct headers (Parameter, Value, Unit, Notes, Last Updated)
  Component 3 (0.30): Assumptions data content matches source — spot-check key rows
  Component 4 (0.20): Original sheets (Revenue, Expenses, Summary) preserved
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_087'


def persist_app_state(domain: str):
    """Best-effort save via Ctrl+S for GUI apps."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    sheet_names = wb.sheetnames

    # Component 1: 'Assumptions' sheet exists in the workbook (0.30 points)
    try:
        if 'Assumptions' in sheet_names:
            print(f"PASS: Component 1 — 'Assumptions' sheet found in workbook (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 — 'Assumptions' sheet not found. Sheets: {sheet_names}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Assumptions sheet has correct headers (0.20 points)
    # Expected headers from source: Parameter, Value, Unit, Notes, Last Updated
    try:
        if 'Assumptions' in sheet_names:
            ws = wb['Assumptions']
            expected_headers = ['Parameter', 'Value', 'Unit', 'Notes', 'Last Updated']
            actual_headers = []
            for col in range(1, 6):
                val = ws.cell(row=1, column=col).value
                actual_headers.append(val)

            matches = sum(1 for a, e in zip(actual_headers, expected_headers)
                         if a is not None and str(a).strip() == e)
            if matches == 5:
                print(f"PASS: Component 2 — All 5 headers match: {actual_headers} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — Headers mismatch. Expected {expected_headers}, got {actual_headers} ({matches}/5 match)")
        else:
            print(f"FAIL: Component 2 — Cannot check headers, 'Assumptions' sheet missing")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Assumptions data content matches source (0.30 points)
    # Spot-check key data rows from the source file
    # Expected data points (row, col, value):
    #   Row 2: Revenue Growth Rate, 0.085
    #   Row 5: Tax Rate, 0.21
    #   Row 11: Terminal Growth Rate, 0.025
    #   Row 14: Accounts Receivable Days, 42
    #   Row 16: Employee Headcount Growth, 0.05
    try:
        if 'Assumptions' in sheet_names:
            ws = wb['Assumptions']
            checks = [
                (2, 1, 'Revenue Growth Rate', 2, 2, 0.085),
                (5, 1, 'Tax Rate', 5, 2, 0.21),
                (11, 1, 'Terminal Growth Rate', 11, 2, 0.025),
                (14, 1, 'Accounts Receivable Days', 14, 2, 42),
                (16, 1, 'Employee Headcount Growth', 16, 2, 0.05),
            ]
            passed_checks = 0
            for name_r, name_c, expected_name, val_r, val_c, expected_val in checks:
                actual_name = ws.cell(row=name_r, column=name_c).value
                actual_val = ws.cell(row=val_r, column=val_c).value
                name_ok = actual_name is not None and str(actual_name).strip() == expected_name
                val_ok = False
                if actual_val is not None:
                    try:
                        val_ok = abs(float(actual_val) - float(expected_val)) < 0.001
                    except (ValueError, TypeError):
                        val_ok = False
                if name_ok and val_ok:
                    passed_checks += 1
                else:
                    print(f"  Data check row {name_r}: name={actual_name}(expect {expected_name}), val={actual_val}(expect {expected_val})")

            # Award proportional credit: 0.30 * (passed / 5)
            comp3_score = 0.30 * (passed_checks / 5)
            if passed_checks == 5:
                print(f"PASS: Component 3 — All 5 data spot-checks passed (0.30 pts)")
                total_score += comp3_score
            elif passed_checks > 0:
                print(f"PARTIAL: Component 3 — {passed_checks}/5 data checks passed ({comp3_score:.2f} pts)")
                total_score += comp3_score
            else:
                print(f"FAIL: Component 3 — 0/5 data checks passed")
        else:
            print(f"FAIL: Component 3 — Cannot check data, 'Assumptions' sheet missing")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Original sheets preserved AND Assumptions sheet added (0.20 points)
    # This is a compound check: task-introduced change (Assumptions added) AND
    # original sheets still intact. Both conditions must hold.
    try:
        required_originals = ['Revenue', 'Expenses', 'Summary']
        originals_present = all(s in sheet_names for s in required_originals)
        assumptions_present = 'Assumptions' in sheet_names

        if not assumptions_present:
            print(f"FAIL: Component 4 — Assumptions sheet not present, compound check fails")
        elif not originals_present:
            missing = [s for s in required_originals if s not in sheet_names]
            print(f"FAIL: Component 4 — Missing original sheets: {missing}")
        else:
            # Verify Revenue sheet still has its header and data
            rev = wb['Revenue']
            rev_a1 = rev.cell(row=1, column=1).value
            rev_a2 = rev.cell(row=2, column=1).value
            rev_b2 = rev.cell(row=2, column=2).value

            data_intact = (
                rev_a1 is not None and str(rev_a1).strip() == 'Product Line'
                and rev_a2 is not None and str(rev_a2).strip() == 'Enterprise Software'
                and rev_b2 is not None and (isinstance(rev_b2, (int, float)) and abs(float(rev_b2) - 1245000) < 1)
            )

            if data_intact:
                print(f"PASS: Component 4 — Original sheets preserved with Assumptions added (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Original sheets exist but Revenue data corrupted. A1={rev_a1}, A2={rev_a2}, B2={rev_b2}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
