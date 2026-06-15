"""
Initial Setup: Insert a sheet from another file using Insert Sheet dialog
Task ID: calc_gsi_087
Domain: libreoffice_calc

Creates:
  1. /home/user/calc_gsi_087.xlsx — main report workbook (Revenue, Expenses sheets)
  2. /home/user/company_assumptions_2024.xlsx — source template with Assumptions sheet
Opens: calc_gsi_087.xlsx in LibreOffice Calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
SOURCE_FILE = f'{WORKDIR}/company_assumptions_2024.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_source_template():
    """Create company_assumptions_2024.xlsx with an Assumptions sheet and other sheets."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Assumptions ---
    ws_assumptions = wb.active
    ws_assumptions.title = 'Assumptions'

    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = ['Parameter', 'Value', 'Unit', 'Notes', 'Last Updated']
    for col, h in enumerate(headers, 1):
        cell = ws_assumptions.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    assumptions_data = [
        ['Revenue Growth Rate', 0.085, '%', 'Based on 3-year CAGR', '2024-01-15'],
        ['Cost of Goods Sold', 0.62, '%', 'Industry average benchmark', '2024-01-15'],
        ['Operating Expense Ratio', 0.18, '%', 'Target from board', '2024-02-01'],
        ['Tax Rate', 0.21, '%', 'Federal corporate rate', '2024-01-10'],
        ['Discount Rate (WACC)', 0.094, '%', 'Updated Q4 2023', '2024-01-20'],
        ['Inflation Rate', 0.032, '%', 'CPI forecast', '2024-01-12'],
        ['Depreciation Rate', 0.15, '%', 'Straight-line method', '2024-01-15'],
        ['Working Capital % Revenue', 0.12, '%', 'Historical average', '2024-02-05'],
        ['Capex % Revenue', 0.08, '%', 'Management guidance', '2024-01-25'],
        ['Terminal Growth Rate', 0.025, '%', 'Long-term GDP growth', '2024-01-15'],
        ['Debt/Equity Ratio', 0.45, 'ratio', 'Target capital structure', '2024-02-01'],
        ['Interest Rate on Debt', 0.055, '%', 'Weighted average cost', '2024-01-18'],
        ['Accounts Receivable Days', 42, 'days', 'Industry benchmark', '2024-01-22'],
        ['Inventory Turnover', 8.5, 'turns', '2023 actual performance', '2024-02-03'],
        ['Employee Headcount Growth', 0.05, '%', 'Hiring plan approved', '2024-01-30'],
    ]

    for r, row_data in enumerate(assumptions_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_assumptions.cell(row=r, column=c, value=val)
            if c == 3 and val == '%':
                ws_assumptions.cell(row=r, column=2).number_format = '0.0%'

    ws_assumptions.column_dimensions['A'].width = 28
    ws_assumptions.column_dimensions['B'].width = 14
    ws_assumptions.column_dimensions['C'].width = 10
    ws_assumptions.column_dimensions['D'].width = 30
    ws_assumptions.column_dimensions['E'].width = 16

    # --- Sheet 2: Version History ---
    ws_history = wb.create_sheet('Version History')
    hist_headers = ['Version', 'Date', 'Author', 'Change Description']
    for col, h in enumerate(hist_headers, 1):
        cell = ws_history.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    hist_data = [
        ['1.0', '2024-01-10', 'Jennifer Walsh', 'Initial assumptions set for FY2024'],
        ['1.1', '2024-01-20', 'Robert Kim', 'Updated WACC after Q4 earnings'],
        ['1.2', '2024-02-01', 'Jennifer Walsh', 'Revised OpEx ratio per board directive'],
        ['1.3', '2024-02-05', 'David Park', 'Added working capital assumptions'],
    ]
    for r, row_data in enumerate(hist_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_history.cell(row=r, column=c, value=val)

    # --- Sheet 3: Methodology ---
    ws_method = wb.create_sheet('Methodology')
    ws_method['A1'] = 'Assumption Methodology Notes'
    ws_method['A1'].font = Font(size=14, bold=True)
    ws_method['A3'] = 'All financial assumptions are reviewed quarterly by the Finance Committee.'
    ws_method['A4'] = 'Revenue growth is based on a 3-year compound annual growth rate analysis.'
    ws_method['A5'] = 'Cost benchmarks are sourced from industry reports (McKinsey, Deloitte).'
    ws_method['A6'] = 'Tax rate reflects current federal corporate tax legislation.'
    ws_method.column_dimensions['A'].width = 70

    wb.save(SOURCE_FILE)
    print(f'Source template created: {SOURCE_FILE}')


def create_main_workbook():
    """Create calc_gsi_087.xlsx — the main report workbook WITHOUT Assumptions sheet."""
    wb = openpyxl.Workbook()

    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    currency_fmt = '$#,##0'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Sheet 1: Revenue ---
    ws_rev = wb.active
    ws_rev.title = 'Revenue'

    rev_headers = ['Product Line', 'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024', 'FY 2024 Total']
    for col, h in enumerate(rev_headers, 1):
        cell = ws_rev.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    revenue_data = [
        ['Enterprise Software', 1245000, 1312000, 1389000, 1456000],
        ['Cloud Services', 892000, 967000, 1045000, 1123000],
        ['Professional Services', 534000, 489000, 567000, 612000],
        ['Hardware Solutions', 345000, 378000, 356000, 401000],
        ['Maintenance & Support', 678000, 695000, 712000, 734000],
        ['Training & Certification', 123000, 145000, 156000, 178000],
        ['Data Analytics Platform', 456000, 512000, 589000, 645000],
        ['Security Solutions', 267000, 298000, 334000, 367000],
        ['Mobile Applications', 189000, 212000, 234000, 256000],
        ['Integration Services', 312000, 334000, 356000, 378000],
        ['Consulting', 445000, 423000, 467000, 489000],
        ['Licensing Fees', 567000, 578000, 590000, 612000],
    ]

    for r, row_data in enumerate(revenue_data, 2):
        ws_rev.cell(row=r, column=1, value=row_data[0]).border = thin_border
        for c, val in enumerate(row_data[1:], 2):
            cell = ws_rev.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
            cell.border = thin_border
        # FY Total formula
        cell = ws_rev.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.font = Font(bold=True)

    # Totals row
    total_row = len(revenue_data) + 2
    ws_rev.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True)
    ws_rev.cell(row=total_row, column=1).border = thin_border
    for c in range(2, 7):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(c)
        cell = ws_rev.cell(row=total_row, column=c,
                           value=f'=SUM({col_letter}2:{col_letter}{total_row - 1})')
        cell.number_format = currency_fmt
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    ws_rev.column_dimensions['A'].width = 26
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_rev.column_dimensions[col_letter].width = 16
    ws_rev.freeze_panes = 'A2'

    # --- Sheet 2: Expenses ---
    ws_exp = wb.create_sheet('Expenses')

    exp_headers = ['Category', 'Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024', 'FY 2024 Total']
    for col, h in enumerate(exp_headers, 1):
        cell = ws_exp.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    expense_data = [
        ['Salaries & Benefits', 2145000, 2178000, 2234000, 2289000],
        ['Office Rent & Utilities', 345000, 345000, 345000, 345000],
        ['Technology Infrastructure', 234000, 245000, 256000, 267000],
        ['Marketing & Advertising', 189000, 212000, 198000, 234000],
        ['Travel & Entertainment', 67000, 78000, 89000, 95000],
        ['Research & Development', 456000, 478000, 501000, 523000],
        ['Legal & Professional', 123000, 134000, 128000, 145000],
        ['Insurance', 89000, 89000, 89000, 89000],
        ['Depreciation', 178000, 178000, 178000, 178000],
        ['Supplies & Equipment', 56000, 62000, 58000, 67000],
        ['Training & Development', 34000, 45000, 38000, 52000],
        ['Miscellaneous', 23000, 28000, 25000, 31000],
    ]

    for r, row_data in enumerate(expense_data, 2):
        ws_exp.cell(row=r, column=1, value=row_data[0]).border = thin_border
        for c, val in enumerate(row_data[1:], 2):
            cell = ws_exp.cell(row=r, column=c, value=val)
            cell.number_format = currency_fmt
            cell.border = thin_border
        cell = ws_exp.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.font = Font(bold=True)

    total_row_exp = len(expense_data) + 2
    ws_exp.cell(row=total_row_exp, column=1, value='TOTAL').font = Font(bold=True)
    ws_exp.cell(row=total_row_exp, column=1).border = thin_border
    for c in range(2, 7):
        col_letter = get_column_letter(c)
        cell = ws_exp.cell(row=total_row_exp, column=c,
                           value=f'=SUM({col_letter}2:{col_letter}{total_row_exp - 1})')
        cell.number_format = currency_fmt
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.fill = PatternFill(start_color='FFD9E2F3', end_color='FFD9E2F3', fill_type='solid')

    ws_exp.column_dimensions['A'].width = 26
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_exp.column_dimensions[col_letter].width = 16
    ws_exp.freeze_panes = 'A2'

    # --- Sheet 3: Summary ---
    ws_sum = wb.create_sheet('Summary')
    ws_sum['A1'] = 'Q4 2024 Financial Summary'
    ws_sum['A1'].font = Font(size=14, bold=True)
    ws_sum['A3'] = 'Report prepared by: Finance Department'
    ws_sum['A4'] = 'Date: January 2025'
    ws_sum['A5'] = 'Status: Draft - Pending Assumptions Review'
    ws_sum['A7'] = 'Note: This workbook requires the company standard assumptions sheet'
    ws_sum['A8'] = 'to be inserted from the central template file for final review.'
    ws_sum['A7'].font = Font(italic=True, color='FF0000')
    ws_sum['A8'].font = Font(italic=True, color='FF0000')
    ws_sum.column_dimensions['A'].width = 55

    wb.save(OUTPUT)
    print(f'Main workbook created: {OUTPUT}')


def main():
    create_source_template()
    create_main_workbook()

    # GUI-ready: open the main workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


main()
