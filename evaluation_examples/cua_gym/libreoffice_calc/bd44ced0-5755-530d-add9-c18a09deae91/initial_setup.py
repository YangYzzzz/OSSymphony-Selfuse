"""
Initial Setup: Create workbook with Annual summary sheet and 12 monthly sales sheets
Task ID: calc_mcp_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

MONTH_FULL = ['January', 'February', 'March', 'April', 'May', 'June',
              'July', 'August', 'September', 'October', 'November', 'December']

# Realistic product data for monthly sheets
PRODUCTS = [
    'Wireless Keyboard', 'USB-C Hub', 'Laptop Stand', 'Monitor Arm',
    'Noise-Cancel Headphones', 'Webcam HD Pro', 'Ergonomic Mouse',
    'Desk Lamp LED', 'Cable Management Kit', 'Portable Charger',
    'Bluetooth Speaker', 'Screen Protector Pack', 'Phone Mount',
    'HDMI Cable 6ft', 'Surge Protector', 'Mousepad XL',
    'USB Flash Drive 64GB', 'Ethernet Adapter', 'Stylus Pen',
    'Laptop Sleeve 15in', 'Desk Organizer', 'Wrist Rest Pad',
    'Mini Projector', 'Smart Power Strip'
]

# Base prices per product (index-aligned with PRODUCTS)
BASE_PRICES = [
    49.99, 34.99, 59.99, 79.99,
    149.99, 89.99, 39.99,
    27.99, 19.99, 24.99,
    64.99, 12.99, 15.99,
    9.99, 29.99, 18.99,
    14.99, 22.99, 29.99,
    34.99, 21.99, 16.99,
    199.99, 44.99
]

# Quantity sold per product per month (varies by month index to create seasonal patterns)
import random
random.seed(42)

def get_qty(month_idx, product_idx):
    """Generate realistic-looking quantities with seasonal variation."""
    base = 10 + (product_idx * 3) % 20
    seasonal = [0.8, 0.75, 0.9, 1.0, 1.05, 1.1, 1.0, 0.95, 1.1, 1.15, 1.3, 1.5]
    qty = int(base * seasonal[month_idx] + ((month_idx * 7 + product_idx * 13) % 11) - 5)
    return max(2, qty)


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Annual sheet (first sheet) ---
    ws_annual = wb.active
    ws_annual.title = 'Annual'

    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    ws_annual.cell(row=1, column=1, value='Month').font = header_font_white
    ws_annual.cell(row=1, column=1).fill = header_fill
    ws_annual.cell(row=1, column=2, value='Total Sales').font = header_font_white
    ws_annual.cell(row=1, column=2).fill = header_fill

    ws_annual.column_dimensions['A'].width = 15
    ws_annual.column_dimensions['B'].width = 18

    for i, month_name in enumerate(MONTH_FULL):
        ws_annual.cell(row=i + 2, column=1, value=month_name)
    # B2:B13 intentionally left EMPTY (task is to add formulas here)

    # --- Create 12 monthly sheets ---
    for m_idx, month_abbr in enumerate(MONTHS):
        ws = wb.create_sheet(month_abbr)

        # Headers: A=Product, B=Category, C=Unit Price, D=Qty Sold, E=Discount%, F=Total
        headers = ['Product', 'Category', 'Unit Price', 'Qty Sold', 'Discount %', 'Total']
        categories = ['Peripherals', 'Accessories', 'Furniture', 'Furniture',
                      'Audio', 'Video', 'Peripherals',
                      'Lighting', 'Accessories', 'Power',
                      'Audio', 'Accessories', 'Accessories',
                      'Cables', 'Power', 'Peripherals',
                      'Storage', 'Networking', 'Peripherals',
                      'Bags', 'Organization', 'Ergonomics',
                      'Video', 'Power']

        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14

        # Data rows 2-25 (24 products)
        for p_idx in range(24):
            row = p_idx + 2
            price = BASE_PRICES[p_idx]
            qty = get_qty(m_idx, p_idx)
            discount = round(((m_idx + p_idx) % 5) * 0.05, 2)  # 0%, 5%, 10%, 15%, 20%
            total = round(price * qty * (1 - discount), 2)

            ws.cell(row=row, column=1, value=PRODUCTS[p_idx])
            ws.cell(row=row, column=2, value=categories[p_idx])
            ws.cell(row=row, column=3, value=price)
            ws.cell(row=row, column=3).number_format = '$#,##0.00'
            ws.cell(row=row, column=4, value=qty)
            ws.cell(row=row, column=5, value=discount)
            ws.cell(row=row, column=5).number_format = '0%'
            ws.cell(row=row, column=6, value=total)
            ws.cell(row=row, column=6).number_format = '$#,##0.00'

        # Row 25: Total row with SUM in F25
        ws.cell(row=25, column=1, value='TOTAL')
        ws.cell(row=25, column=1).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=25, column=6, value=f'=SUM(F2:F24)')
        ws.cell(row=25, column=6).font = Font(name='Calibri', size=11, bold=True)
        ws.cell(row=25, column=6).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
