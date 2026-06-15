"""
Reward Script: Cross-sheet formula references in Annual summary
Task ID: calc_mcp_040
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 12 cells B2:B13 contain formulas (non-empty, string starting with '=')
  Component 2 (0.5): Each formula correctly references the corresponding month sheet cell F25
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_040'

# Mapping: row -> expected sheet name prefix in formula
MONTH_MAP = {
    2: 'Jan',
    3: 'Feb',
    4: 'Mar',
    5: 'Apr',
    6: 'May',
    7: 'Jun',
    8: 'Jul',
    9: 'Aug',
    10: 'Sep',
    11: 'Oct',
    12: 'Nov',
    13: 'Dec',
}


def persist_app_state():
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def normalize_formula(formula_str):
    """Normalize a formula for comparison: uppercase, strip spaces."""
    if not isinstance(formula_str, str):
        return ""
    return formula_str.upper().replace(" ", "")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check Annual sheet exists
    if 'Annual' not in wb.sheetnames:
        print("FAIL: 'Annual' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Annual']

    # Component 1: All 12 cells B2:B13 contain formulas (0.5 points)
    # Award partial: each cell with a formula contributes 0.5/12 points
    try:
        formula_count = 0
        for row in range(2, 14):
            val = ws.cell(row=row, column=2).value
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1
            else:
                print(f"FAIL: B{row} does not contain a formula (value: {val})")

        comp1_score = (formula_count / 12.0) * 0.5
        if formula_count == 12:
            print(f"PASS: Component 1 - All 12 cells B2:B13 contain formulas (0.5 pts)")
            total_score += comp1_score
        elif formula_count > 0:
            print(f"PARTIAL: Component 1 - {formula_count}/12 cells contain formulas ({comp1_score:.3f} pts)")
            total_score += comp1_score
        else:
            print(f"FAIL: Component 1 - No cells contain formulas (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Each formula correctly references the corresponding month sheet F25 (0.5 points)
    # Award partial: each correct reference contributes 0.5/12 points
    # Accept various valid formula patterns:
    #   =Jan.F25, =Jan!F25, ='Jan'.F25, ='Jan'!F25 (with or without quotes, . or !)
    try:
        correct_count = 0
        for row, month in MONTH_MAP.items():
            val = ws.cell(row=row, column=2).value
            if not isinstance(val, str):
                print(f"FAIL: B{row} is not a formula, cannot check reference")
                continue

            norm = normalize_formula(val)
            # Accept patterns like =JAN.F25, =JAN!F25, ='JAN'.F25, ='JAN'!F25
            month_upper = month.upper()
            pattern = re.compile(
                r"^='{0,1}" + month_upper + r"'{0,1}[.!]F25$"
            )
            if pattern.match(norm):
                correct_count += 1
            else:
                print(f"FAIL: B{row} formula '{val}' does not match expected reference to {month}!F25")

        comp2_score = (correct_count / 12.0) * 0.5
        if correct_count == 12:
            print(f"PASS: Component 2 - All 12 formulas reference correct month sheet F25 (0.5 pts)")
            total_score += comp2_score
        elif correct_count > 0:
            print(f"PARTIAL: Component 2 - {correct_count}/12 correct references ({comp2_score:.3f} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 - No correct references found (0.0 pts)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
