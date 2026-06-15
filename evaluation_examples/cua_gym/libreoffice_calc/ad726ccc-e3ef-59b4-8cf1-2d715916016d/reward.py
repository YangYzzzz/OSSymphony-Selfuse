"""
Reward Script: Apply region-based discounts using IFS formulas in column C
Task ID: calc_fma_ifs_region_discount_079
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 14 cells C2:C15 contain IFS formulas
  Component 2 (0.5): Formulas have correct discount factors for all 5 regions
                     AND no unintended changes to columns A/B or header C1
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_ifs_region_discount_079'
SHEET_NAME = 'RegionalPricing'

# Expected discount multipliers per region
EXPECTED_DISCOUNTS = {
    'North': 0.9,
    'South': 0.92,
    'East': 0.88,
    'West': 0.95,
    'International': 0.85,
}

# Expected region data in column A (rows 2-15)
EXPECTED_REGIONS = [
    'North', 'South', 'East', 'West', 'International',
    'North', 'East', 'South', 'International', 'West',
    'North', 'South', 'East', 'West'
]

# Expected base prices in column B (rows 2-15)
EXPECTED_PRICES = [
    100, 250, 80, 320, 500,
    150, 90, 180, 420, 275,
    110, 200, 75, 340
]


def check_ifs_formula_discounts(formula, row_num):
    """
    Check if an IFS formula contains the correct discount factors for all 5 regions.
    Returns a set of region names whose discount factors are correctly encoded.
    """
    correct_regions = set()
    if not isinstance(formula, str):
        return correct_regions

    formula_upper = formula.upper().replace(' ', '')

    # Check each region's discount factor
    discount_patterns = {
        'North': ['B{}*0.9'.format(row_num), 'B{}*0.90'.format(row_num)],
        'South': ['B{}*0.92'.format(row_num)],
        'East': ['B{}*0.88'.format(row_num)],
        'West': ['B{}*0.95'.format(row_num)],
        'International': ['B{}*0.85'.format(row_num)],
    }

    for region, patterns in discount_patterns.items():
        for pat in patterns:
            if pat.upper() in formula_upper:
                correct_regions.add(region)
                break

    return correct_regions


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: check the sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -------------------------------------------------------------------------
    # Component 1: IFS formulas present in all 14 cells C2:C15 (0.5 points)
    # Verifies that the agent populated the previously-empty C2:C15 range
    # with IFS formulas (changes from None to formula string).
    # -------------------------------------------------------------------------
    try:
        cells_with_ifs = 0
        for row in range(2, 16):
            val = ws.cell(row=row, column=3).value
            if isinstance(val, str) and val.upper().replace(' ', '').startswith('=IFS('):
                cells_with_ifs += 1

        if cells_with_ifs == 14:
            print(f"PASS: Component 1 — All 14 cells C2:C15 contain IFS formulas (0.5 pts)")
            total_score += 0.5
        elif cells_with_ifs > 0:
            partial = round(0.5 * cells_with_ifs / 14, 3)
            print(f"PARTIAL: Component 1 — {cells_with_ifs}/14 cells have IFS formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No IFS formulas found in C2:C15 (cells still empty or wrong formula type)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Correct discount factors AND no unintended modifications (0.5 pts)
    # Only awards points when formulas are present AND correct (compound check).
    # This prevents partial credit for the pre-existing unchanged state.
    # -------------------------------------------------------------------------
    try:
        formulas_correct = True
        correct_formula_count = 0

        # First, check that C2:C15 has formulas — this is required for Component 2 to pass
        any_ifs = any(
            isinstance(ws.cell(row=row, column=3).value, str) and
            ws.cell(row=row, column=3).value.upper().replace(' ', '').startswith('=IFS(')
            for row in range(2, 16)
        )

        if not any_ifs:
            print(f"FAIL: Component 2 — No IFS formulas present; cannot verify correctness")
        else:
            # Verify all 14 rows have correct discount factors
            for idx, row in enumerate(range(2, 16)):
                val = ws.cell(row=row, column=3).value
                if not isinstance(val, str) or not val.upper().replace(' ', '').startswith('=IFS('):
                    formulas_correct = False
                    continue

                correct_regions = check_ifs_formula_discounts(val, row)
                if len(correct_regions) == 5:
                    correct_formula_count += 1
                else:
                    formulas_correct = False
                    missing = set(EXPECTED_DISCOUNTS.keys()) - correct_regions
                    print(f"  FAIL in C{row}: formula missing correct discount for: {missing}")

            # Check no unintended changes to A/B columns and C1 header
            data_intact = True
            data_issues = []

            c1_val = ws.cell(row=1, column=3).value
            if c1_val != 'Discounted Price':
                data_intact = False
                data_issues.append(f"C1 header: expected 'Discounted Price', found {repr(c1_val)}")

            for idx, row in enumerate(range(2, 16)):
                a_val = ws.cell(row=row, column=1).value
                b_val = ws.cell(row=row, column=2).value
                if a_val != EXPECTED_REGIONS[idx]:
                    data_intact = False
                    data_issues.append(f"A{row}: expected {repr(EXPECTED_REGIONS[idx])}, found {repr(a_val)}")
                if b_val != EXPECTED_PRICES[idx]:
                    data_intact = False
                    data_issues.append(f"B{row}: expected {repr(EXPECTED_PRICES[idx])}, found {repr(b_val)}")

            if correct_formula_count == 14 and data_intact:
                print(f"PASS: Component 2 — All 14 IFS formulas have correct discount factors "
                      f"and columns A/B/C1 header are intact (0.5 pts)")
                total_score += 0.5
            elif correct_formula_count > 0 and data_intact:
                partial = round(0.5 * correct_formula_count / 14, 3)
                print(f"PARTIAL: Component 2 — {correct_formula_count}/14 correct formulas, "
                      f"data intact ({partial} pts)")
                total_score += partial
            elif correct_formula_count == 14 and not data_intact:
                print(f"PARTIAL: Component 2 — Formulas correct but unintended changes found:")
                for issue in data_issues[:3]:
                    print(f"  - {issue}")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — {correct_formula_count}/14 correct formulas. "
                      f"Data issues: {data_issues[:2] if data_issues else 'none'}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
