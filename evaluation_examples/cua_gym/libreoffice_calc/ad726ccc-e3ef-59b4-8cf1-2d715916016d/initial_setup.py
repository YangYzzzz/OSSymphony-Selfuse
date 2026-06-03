"""
Initial Setup: Region-based discount spreadsheet (IFS formula task)
Task ID: calc_fma_ifs_region_discount_079
Domain: libreoffice_calc

Creates RegionalPricing sheet with Region (col A) and Base Price (col B) data.
Column C has the header 'Discounted Price' but cells C2:C15 are left empty.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_ifs_region_discount_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'RegionalPricing'

    # --- Row 1: Headers ---
    ws['A1'] = 'Region'
    ws['B1'] = 'Base Price'
    ws['C1'] = 'Discounted Price'

    # --- Rows 2-15: Region names and base prices (from task context) ---
    regions = [
        'North', 'South', 'East', 'West', 'International',
        'North', 'East', 'South', 'International', 'West',
        'North', 'South', 'East', 'West'
    ]
    prices = [
        100.00, 250.00, 80.00, 320.00, 500.00,
        150.00, 90.00, 180.00, 420.00, 275.00,
        110.00, 200.00, 75.00, 340.00
    ]

    for i, (region, price) in enumerate(zip(regions, prices), start=2):
        ws.cell(row=i, column=1, value=region)
        ws.cell(row=i, column=2, value=price)
        # Column C intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
