"""
Initial Setup: Compute employee tenure in years and months using DATEDIF
Task ID: osworld_calc_age_calculation_datedif_007
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_age_calculation_datedif_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: HR Data ---
    ws = wb.active
    ws.title = 'HR Data'

    # Headers
    headers = ['Employee ID', 'Hire Date', 'Tenure', 'Tenure Category']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # Employee data (realistic names, varied hire dates giving different tenures)
    # Hire dates are chosen to produce a realistic spread across Junior, Developing, Experienced, Senior
    data = [
        ('EMP001', '2024-08-15'),   # ~0.5 years -> Junior
        ('EMP002', '2023-11-01'),   # ~1.3 years -> Developing
        ('EMP003', '2022-05-20'),   # ~2.8 years -> Developing
        ('EMP004', '2021-03-10'),   # ~4 years -> Experienced
        ('EMP005', '2019-07-22'),   # ~6.5 years -> Senior
        ('EMP006', '2018-01-05'),   # ~8 years -> Senior
        ('EMP007', '2024-01-30'),   # ~1.1 years -> Developing
        ('EMP008', '2020-09-14'),   # ~4.4 years -> Experienced
        ('EMP009', '2016-04-03'),   # ~9.9 years -> Senior
        ('EMP010', '2023-06-17'),   # ~1.7 years -> Developing
        ('EMP011', '2021-11-28'),   # ~3.3 years -> Experienced
        ('EMP012', '2025-01-10'),   # ~0.2 years -> Junior
        ('EMP013', '2017-08-09'),   # ~7.5 years -> Senior
        ('EMP014', '2022-12-01'),   # ~2.2 years -> Developing
        ('EMP015', '2020-03-25'),   # ~4.9 years -> Experienced
    ]

    from datetime import date
    for r, (emp_id, hire_date_str) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=emp_id)
        # Store hire date as a date object so LibreOffice recognizes it as a date
        hire_date = date.fromisoformat(hire_date_str)
        hire_cell = ws.cell(row=r, column=2, value=hire_date)
        hire_cell.number_format = 'yyyy-mm-dd'
        # Columns C and D are intentionally left empty (task requires filling them)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
