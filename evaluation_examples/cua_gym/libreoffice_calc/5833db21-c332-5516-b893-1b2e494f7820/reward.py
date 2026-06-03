"""
Reward Script: Compute employee tenure in years and months, add tenure category column,
               and highlight Senior employees in green.
Task ID: osworld_calc_age_calculation_datedif_007
Domain: libreoffice_calc
Scoring:
  Component 1: Column C has DATEDIF tenure formulas (0.35 pts)
  Component 2: Column D has IFS/nested-IF category formulas (0.35 pts)
  Component 3: Conditional formatting highlights Senior rows in green (0.30 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_age_calculation_datedif_007'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Scoring rubric:
    - 0.35 pts: Column C filled with DATEDIF tenure formulas (e.g. '=DATEDIF(B2,TODAY(),"Y")...')
    - 0.35 pts: Column D filled with IFS/IF category formulas containing Junior/Developing/Experienced/Senior
    - 0.30 pts: Conditional formatting rule applied with green fill for Senior rows
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify the expected sheet exists
    if 'HR Data' not in wb.sheetnames:
        print("CRITICAL: Sheet 'HR Data' not found. Cannot score.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['HR Data']

    # Determine data rows (rows 2 through max_row)
    max_row = ws.max_row
    data_rows = list(range(2, max_row + 1))
    if len(data_rows) == 0:
        print("CRITICAL: No data rows found.")
        print("REWARD: 0.0")
        return 0.0

    # ------------------------------------------------------------------
    # Component 1: Column C has DATEDIF tenure formulas (0.35 points)
    # Verify that each data row in column C contains a formula that uses
    # DATEDIF and produces a formatted tenure string (e.g., 'X years Y months').
    # ------------------------------------------------------------------
    try:
        c_formula_count = 0
        c_datedif_count = 0
        c_tenure_format_count = 0

        for r in data_rows:
            cell_c = ws.cell(row=r, column=3)
            val = cell_c.value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                c_formula_count += 1
                val_upper = val.upper()
                if 'DATEDIF' in val_upper:
                    c_datedif_count += 1
                # Check for both year and month components (for formatted tenure string)
                # pattern: "Y" for years and "YM" for months remaining
                if '"Y"' in val_upper and '"YM"' in val_upper:
                    c_tenure_format_count += 1

        print(f"  Column C: {c_formula_count}/{len(data_rows)} cells have formulas")
        print(f"  Column C: {c_datedif_count}/{len(data_rows)} cells use DATEDIF")
        print(f"  Column C: {c_tenure_format_count}/{len(data_rows)} cells have year+month tenure format")

        if c_tenure_format_count == len(data_rows) and c_datedif_count == len(data_rows):
            print(f"PASS: Component 1 — All {len(data_rows)} rows in column C have DATEDIF tenure formulas (0.35 pts)")
            total_score += 0.35
        elif c_datedif_count == len(data_rows):
            # Has DATEDIF but maybe not the exact format - partial credit
            print(f"PASS (partial): Component 1 — All rows use DATEDIF but tenure format incomplete (0.20 pts)")
            total_score += 0.20
        elif c_formula_count == len(data_rows) and c_datedif_count > 0:
            print(f"FAIL: Component 1 — Formulas present but DATEDIF not used for all rows ({c_datedif_count}/{len(data_rows)})")
        elif c_formula_count > 0:
            print(f"FAIL: Component 1 — Only {c_formula_count}/{len(data_rows)} rows have formulas in column C")
        else:
            print(f"FAIL: Component 1 — Column C is empty, no tenure formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Column D has IFS/nested-IF category formulas (0.35 points)
    # Verify that each data row in column D contains a formula that assigns
    # one of: Junior, Developing, Experienced, Senior based on years of service.
    # ------------------------------------------------------------------
    try:
        d_formula_count = 0
        d_category_count = 0
        categories = {'JUNIOR', 'DEVELOPING', 'EXPERIENCED', 'SENIOR'}

        for r in data_rows:
            cell_d = ws.cell(row=r, column=4)
            val = cell_d.value
            if val is not None and isinstance(val, str) and val.strip().startswith('='):
                d_formula_count += 1
                val_upper = val.upper()
                # Check that all four categories are present in the formula
                if all(cat in val_upper for cat in categories):
                    d_category_count += 1

        print(f"  Column D: {d_formula_count}/{len(data_rows)} cells have formulas")
        print(f"  Column D: {d_category_count}/{len(data_rows)} cells reference all 4 categories")

        if d_category_count == len(data_rows):
            print(f"PASS: Component 2 — All {len(data_rows)} rows in column D have correct category formulas (0.35 pts)")
            total_score += 0.35
        elif d_formula_count == len(data_rows):
            # Has formulas but not all 4 categories
            missing_cats_count = sum(
                1 for r in data_rows
                if ws.cell(row=r, column=4).value and isinstance(ws.cell(row=r, column=4).value, str)
                and not all(cat in ws.cell(row=r, column=4).value.upper() for cat in categories)
            )
            print(f"FAIL: Component 2 — {d_formula_count} formulas present but only {d_category_count} contain all 4 categories")
        elif d_formula_count > 0:
            print(f"FAIL: Component 2 — Only {d_formula_count}/{len(data_rows)} rows have formulas in column D")
        else:
            print(f"FAIL: Component 2 — Column D is empty, no category formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Conditional formatting highlights Senior rows in green (0.30 points)
    # Verify a conditional formatting rule exists that:
    #   - Uses a formula referencing "Senior" in column D
    #   - Applies a green fill color
    # ------------------------------------------------------------------
    try:
        cf_found = False
        green_fill_found = False
        senior_condition_found = False

        # Green color patterns (various shades of green commonly used)
        # FF92D050 is lime green, FF00FF00 is pure green, also check for common variants
        green_patterns = [
            'FF92D050',  # light lime green (Excel default green highlight)
            'FF00FF00',  # pure green
            'FF70AD47',  # medium green
            'FF00B050',  # darker green
            'FF92D050'.lower(),
        ]

        for cf_range in ws.conditional_formatting:
            for rule in ws.conditional_formatting[cf_range]:
                cf_found = True
                # Check if the formula references "Senior" in column D
                if hasattr(rule, 'formula') and rule.formula:
                    for formula in rule.formula:
                        if formula and 'SENIOR' in str(formula).upper():
                            senior_condition_found = True

                # Check fill color is green
                if hasattr(rule, 'dxf') and rule.dxf:
                    dxf = rule.dxf
                    if dxf.fill:
                        try:
                            fill_color = dxf.fill.fgColor.rgb
                            # Check if any component of the RGB is greenish
                            # A "green" color should have G channel significantly higher than R and B
                            if fill_color:
                                fill_upper = fill_color.upper()
                                if fill_upper in [p.upper() for p in green_patterns]:
                                    green_fill_found = True
                                else:
                                    # Parse ARGB components: AARRGGBB
                                    if len(fill_upper) == 8:
                                        try:
                                            r_val = int(fill_upper[2:4], 16)
                                            g_val = int(fill_upper[4:6], 16)
                                            b_val = int(fill_upper[6:8], 16)
                                            # Green: G channel dominant and reasonably high
                                            if g_val > r_val and g_val > b_val and g_val >= 100:
                                                green_fill_found = True
                                        except ValueError:
                                            pass
                        except Exception:
                            pass

        print(f"  Conditional formatting found: {cf_found}")
        print(f"  Senior condition in formula: {senior_condition_found}")
        print(f"  Green fill applied: {green_fill_found}")

        if senior_condition_found and green_fill_found:
            print(f"PASS: Component 3 — Conditional formatting highlights Senior rows in green (0.30 pts)")
            total_score += 0.30
        elif senior_condition_found and cf_found:
            print(f"FAIL: Component 3 — CF rule references Senior but fill color is not green")
        elif green_fill_found and cf_found:
            print(f"FAIL: Component 3 — CF rule has green fill but formula does not reference Senior")
        elif cf_found:
            print(f"FAIL: Component 3 — CF rule exists but neither Senior condition nor green fill confirmed")
        else:
            print(f"FAIL: Component 3 — No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
