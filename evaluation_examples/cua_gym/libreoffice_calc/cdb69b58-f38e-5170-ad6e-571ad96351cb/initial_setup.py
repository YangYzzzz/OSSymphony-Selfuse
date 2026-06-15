"""
Initial Setup: Create a spreadsheet with monthly sales and profit margin data
Task ID: calc_chart_combo_bar_line_015
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_combo_bar_line_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Performance ---
    ws = wb.active
    ws.title = 'Performance'

    # Headers
    headers = ['Month', 'Sales ($000)', 'Profit Margin %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center')
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    # Data rows (as specified in context)
    data = [
        ['Jan', 320, 18.5],
        ['Feb', 295, 17.2],
        ['Mar', 410, 21.3],
        ['Apr', 385, 20.1],
        ['May', 445, 22.8],
        ['Jun', 490, 24.1],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.alignment = Alignment(horizontal='center')
            elif c == 2:
                cell.number_format = '#,##0'
            elif c == 3:
                cell.number_format = '0.0'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 18

    # NO charts — task requires the agent to create the chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
