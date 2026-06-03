"""
Initial Setup: Create performance analysis spreadsheet with line chart showing only Actual series
Task ID: calc_gg2_030
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_030'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Analysis Sheet ---
    ws = wb.active
    ws.title = 'Analysis'

    # Headers
    headers = ['Month', 'Period', 'Actual', 'Target']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Data: 12 months of actual vs target performance metrics (in thousands)
    data = [
        ['January',    'Q1', 142.5, 150.0],
        ['February',   'Q1', 158.3, 155.0],
        ['March',      'Q1', 167.8, 160.0],
        ['April',      'Q2', 153.2, 165.0],
        ['May',        'Q2', 172.6, 170.0],
        ['June',       'Q2', 189.4, 175.0],
        ['July',       'Q3', 178.1, 180.0],
        ['August',     'Q3', 195.7, 185.0],
        ['September',  'Q3', 201.3, 190.0],
        ['October',    'Q4', 188.9, 195.0],
        ['November',   'Q4', 210.5, 200.0],
        ['December',   'Q4', 225.8, 210.0],
    ]

    data_font = Font(name='Calibri', size=11)
    num_fmt = '#,##0.0'

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c >= 3:  # Actual and Target columns
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal='right')
            elif c == 2:
                cell.alignment = Alignment(horizontal='center')

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12

    # --- Create Line Chart with ONLY Actual series ---
    chart = LineChart()
    chart.title = 'Monthly Performance'
    chart.y_axis.title = 'Value (thousands)'
    chart.x_axis.title = 'Month'
    chart.width = 18
    chart.height = 12

    # Only add Actual series (column C), NOT Target (column D)
    actual_data = Reference(ws, min_col=3, min_row=1, max_row=13)
    categories = Reference(ws, min_col=1, min_row=2, max_row=13)
    chart.add_data(actual_data, titles_from_data=True)
    chart.set_categories(categories)

    # Style the Actual series as a solid blue line
    actual_series = chart.series[0]
    actual_series.graphicalProperties.line.solidFill = '4472C4'
    actual_series.graphicalProperties.line.dashStyle = 'solid'
    actual_series.graphicalProperties.line.width = 22000  # ~1.75pt in EMU

    ws.add_chart(chart, 'F2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
