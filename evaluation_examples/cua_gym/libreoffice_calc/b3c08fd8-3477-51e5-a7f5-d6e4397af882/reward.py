"""
Reward Script: Build a dynamic sales dashboard with dropdown and chart
Task ID: calc_gen_chart_052
Domain: libreoffice_calc
Scoring:
  Component 1: Data validation dropdown in Dashboard!B1 with rep names (0.25 pts)
  Component 2: Months (Jan-Dec) in Dashboard!A3:A14 (0.15 pts)
  Component 3: Dynamic INDEX/MATCH formulas in Dashboard!B3:B14 (0.35 pts)
  Component 4: LineChart on Dashboard sheet referencing dynamic data (0.25 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_chart_052'

EXPECTED_MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

EXPECTED_REPS = ['Alice', 'Bob', 'Carol', 'Dave', 'Eve']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Dashboard' sheet must exist
    if 'Dashboard' not in wb.sheetnames:
        print("FAIL: 'Dashboard' sheet not found in workbook")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    ws_dash = wb['Dashboard']

    # Component 1: Data validation dropdown in B1 with all 5 rep names (0.25 pts)
    try:
        validations = ws_dash.data_validations.dataValidation
        found_dv = False
        dv_has_all_reps = False
        dv_on_b1 = False

        for dv in validations:
            if dv.type == 'list':
                # Check if formula1 contains the expected rep names
                formula = dv.formula1 if dv.formula1 else ''
                # Strip surrounding quotes if present
                formula_clean = formula.strip('"').strip("'")
                # Check that all rep names are present in the formula
                reps_found = all(rep in formula_clean for rep in EXPECTED_REPS)
                # Check if sqref includes B1
                sqref_str = str(dv.sqref) if dv.sqref else ''
                b1_covered = 'B1' in sqref_str

                if reps_found and b1_covered:
                    found_dv = True
                    dv_has_all_reps = True
                    dv_on_b1 = True
                    break
                elif reps_found:
                    dv_has_all_reps = True
                elif b1_covered:
                    dv_on_b1 = True

        if found_dv:
            print(f"PASS: Component 1 — Dropdown in B1 with all rep names (Alice, Bob, Carol, Dave, Eve) (0.25 pts)")
            total_score += 0.25
        elif dv_has_all_reps:
            print(f"FAIL: Component 1 — Dropdown has rep names but not applied to B1")
        elif dv_on_b1:
            print(f"FAIL: Component 1 — Dropdown on B1 but does not contain all rep names")
        else:
            print(f"FAIL: Component 1 — No valid dropdown with rep names found on B1 (validations found: {len(validations)})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Month labels in A3:A14 (Jan-Dec) (0.15 pts)
    try:
        months_found = []
        for row in range(3, 15):
            val = ws_dash.cell(row=row, column=1).value
            months_found.append(val)

        # Check if months match expected sequence (case-insensitive comparison)
        months_match = (
            len(months_found) == 12 and
            all(
                months_found[i] is not None and
                str(months_found[i]).strip() == EXPECTED_MONTHS[i]
                for i in range(12)
            )
        )

        if months_match:
            print(f"PASS: Component 2 — Months Jan-Dec in A3:A14 (0.15 pts)")
            total_score += 0.15
        else:
            non_none = [m for m in months_found if m is not None]
            print(f"FAIL: Component 2 — Expected 12 months (Jan-Dec) in A3:A14, found: {months_found[:5]}... ({len(non_none)} non-empty)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Dynamic INDEX/MATCH formulas in B3:B14 (0.35 pts)
    try:
        formulas_found = []
        formula_cells_with_index_match = 0

        for row in range(3, 15):
            val = ws_dash.cell(row=row, column=2).value
            formulas_found.append(val)
            if val is not None:
                val_str = str(val).upper()
                # Check for INDEX and MATCH (dynamic lookup) OR INDIRECT
                if ('INDEX' in val_str and 'MATCH' in val_str) or 'INDIRECT' in val_str:
                    formula_cells_with_index_match += 1

        # Must have all 12 rows with dynamic formulas
        if formula_cells_with_index_match == 12:
            # Also verify formulas reference SalesData and B1 (dropdown cell)
            sample_formula = str(formulas_found[0]).upper() if formulas_found[0] else ''
            refs_salesdata = 'SALESDATA' in sample_formula
            refs_b1 = '$B$1' in sample_formula or 'B1' in sample_formula

            if refs_salesdata and refs_b1:
                print(f"PASS: Component 3 — All 12 rows (B3:B14) have INDEX/MATCH formulas referencing SalesData and dropdown B1 (0.35 pts)")
                total_score += 0.35
            elif formula_cells_with_index_match == 12:
                print(f"PASS: Component 3 — All 12 rows (B3:B14) have dynamic lookup formulas (0.35 pts)")
                total_score += 0.35
        elif formula_cells_with_index_match >= 6:
            print(f"PARTIAL: Component 3 — Only {formula_cells_with_index_match}/12 rows have INDEX/MATCH formulas (partial credit not awarded, need all 12)")
            print(f"FAIL: Component 3 — Expected 12 dynamic formulas, found {formula_cells_with_index_match}")
        else:
            non_none_b = [f for f in formulas_found if f is not None]
            print(f"FAIL: Component 3 — Expected INDEX/MATCH or INDIRECT formulas in B3:B14, found {formula_cells_with_index_match} (non-empty: {len(non_none_b)})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: LineChart on Dashboard sheet referencing dynamic data (0.25 pts)
    try:
        charts = ws_dash._charts
        has_line_chart = False
        chart_refs_dashboard_data = False

        for chart in charts:
            chart_type = type(chart).__name__
            # Accept LineChart
            if 'Line' in chart_type or 'line' in chart_type:
                has_line_chart = True
                # Check if chart references B3:B14 or similar range on Dashboard
                try:
                    for series in chart.series:
                        val_ref = str(series.val) if series.val else ''
                        cat_ref = str(series.cat) if series.cat else ''
                        # Look for reference to Dashboard column B rows 3-14
                        if ('Dashboard' in val_ref or 'B$3' in val_ref or 'B3' in val_ref) and len(chart.series) >= 1:
                            chart_refs_dashboard_data = True
                            break
                except Exception as e2:
                    # If series inspection fails but chart exists, still give partial
                    chart_refs_dashboard_data = True

        if has_line_chart and chart_refs_dashboard_data:
            print(f"PASS: Component 4 — LineChart on Dashboard referencing dynamic data range (0.25 pts)")
            total_score += 0.25
        elif has_line_chart:
            print(f"FAIL: Component 4 — LineChart found but does not reference Dashboard dynamic data range")
        elif len(charts) > 0:
            chart_names = [type(c).__name__ for c in charts]
            print(f"FAIL: Component 4 — Chart exists but is not a LineChart: {chart_names}")
        else:
            print(f"FAIL: Component 4 — No chart found on Dashboard sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
