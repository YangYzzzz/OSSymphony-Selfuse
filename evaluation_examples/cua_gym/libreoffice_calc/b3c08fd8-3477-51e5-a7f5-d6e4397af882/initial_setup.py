"""
Initial Setup: Dynamic sales dashboard with chart based on dropdown selection
Task ID: calc_gen_chart_052
Domain: libreoffice_calc

Creates:
  - Sheet 'SalesData': monthly revenue data for 5 sales reps (Alice, Bob, Carol, Dave, Eve), Jan-Dec
  - Sheet 'Dashboard': empty sheet (task is to build the dashboard with dropdown + dynamic chart)
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_chart_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ── Sheet 1: SalesData ──────────────────────────────────────────────────
    ws_data = wb.active
    ws_data.title = 'SalesData'

    # Headers
    headers = ['Month', 'Alice', 'Bob', 'Carol', 'Dave', 'Eve']
    for col, h in enumerate(headers, 1):
        cell = ws_data.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center')

    # Monthly revenue data — realistic sales figures (in USD)
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # Each rep has a distinct performance profile
    alice_data  = [42500, 38700, 51200, 47800, 55300, 62100,
                   58400, 64200, 59700, 71500, 68300, 79800]
    bob_data    = [35100, 33400, 39200, 41600, 44800, 48300,
                   46500, 52100, 49800, 57200, 61400, 65900]
    carol_data  = [51200, 49800, 56300, 53700, 61200, 67400,
                   63800, 70100, 66500, 74300, 78900, 85600]
    dave_data   = [28600, 31200, 34700, 37500, 40100, 43800,
                   42200, 47600, 45300, 51800, 55200, 59700]
    eve_data    = [39800, 41500, 45200, 48700, 52300, 56900,
                   54100, 60400, 57800, 65200, 69700, 74100]

    rep_data = [alice_data, bob_data, carol_data, dave_data, eve_data]

    for row_idx, month in enumerate(months, 2):
        ws_data.cell(row=row_idx, column=1, value=month)
        for col_idx, rep_vals in enumerate(rep_data, 2):
            ws_data.cell(row=row_idx, column=col_idx, value=rep_vals[row_idx - 2])

    # Column widths for readability
    ws_data.column_dimensions['A'].width = 10
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_data.column_dimensions[col_letter].width = 12

    # Freeze header row
    ws_data.freeze_panes = 'A2'

    # ── Sheet 2: Dashboard (empty — agent must build this) ──────────────────
    ws_dash = wb.create_sheet('Dashboard')
    # Leave intentionally empty; agent will add dropdown, formulas, and chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: SalesData (with Jan-Dec data for Alice/Bob/Carol/Dave/Eve), Dashboard (empty)')


create_initial()
