"""
Initial Setup: Loan amortization summary spreadsheet (pre-task state)
Task ID: calc_fmb_complex_financial_068
Domain: libreoffice_calc

Creates a spreadsheet with loan parameters filled in but B6/B7/B8 left empty
so the agent must enter the PMT-based formulas.
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_complex_financial_068'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ----- Sheet: Amortization Summary -----
    ws = wb.active
    ws.title = 'Amortization Summary'

    # Title row
    ws['A1'] = 'Loan Amortization Summary'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Style helpers
    label_font = Font(name='Calibri', size=11, bold=True)
    value_font = Font(name='Calibri', size=11)
    header_fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Loan parameters (A2:B5) — pre-filled
    params = [
        ('Loan Amount',       200000),
        ('Annual Rate',       0.0525),
        ('Term (Years)',      15),
        ('Monthly Payments',  12),
    ]
    for i, (label, value) in enumerate(params, start=2):
        ws.cell(row=i, column=1, value=label).font = label_font
        ws.cell(row=i, column=1).fill = header_fill
        ws.cell(row=i, column=1).border = border
        ws.cell(row=i, column=1).alignment = Alignment(horizontal='left')

        val_cell = ws.cell(row=i, column=2, value=value)
        val_cell.font = value_font
        val_cell.border = border
        val_cell.alignment = Alignment(horizontal='right')

    # Number formats for parameter values
    ws['B2'].number_format = '$#,##0.00'
    ws['B3'].number_format = '0.00%'
    ws['B4'].number_format = '0'
    ws['B5'].number_format = '0'

    # Calculated rows (A6:A8) — labels present, values EMPTY (agent must fill)
    calc_labels = [
        'Monthly Payment',
        'Total Paid',
        'Total Interest',
    ]
    for i, label in enumerate(calc_labels, start=6):
        lbl_cell = ws.cell(row=i, column=1, value=label)
        lbl_cell.font = label_font
        lbl_cell.fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
        lbl_cell.border = border
        lbl_cell.alignment = Alignment(horizontal='left')

        # Value cell — intentionally left EMPTY
        val_cell = ws.cell(row=i, column=2, value=None)
        val_cell.font = value_font
        val_cell.border = border
        val_cell.number_format = '$#,##0.00'
        val_cell.alignment = Alignment(horizontal='right')

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18

    # Row heights
    ws.row_dimensions[1].height = 28
    for r in range(2, 9):
        ws.row_dimensions[r].height = 20

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('State: B6, B7, B8 are EMPTY (agent must enter PMT formulas)')


create_initial()
