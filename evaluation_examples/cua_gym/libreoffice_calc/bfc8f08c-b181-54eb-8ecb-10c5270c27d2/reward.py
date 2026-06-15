"""
Reward Script: Set up a loan amortization summary with PMT, total paid, and total interest formulas.
Task ID: calc_fmb_complex_financial_068
Domain: libreoffice_calc
Scoring:
  Component 1: B6 contains PMT formula for monthly payment         — 0.4 points
  Component 2: B7 contains formula for total amount paid           — 0.3 points
  Component 3: B8 contains formula for total interest paid         — 0.3 points
  Total: 1.0

Notes:
  - Loan: $200,000 at 5.25% annual rate for 15 years
  - B6 = =PMT(B3/B5,B4*B5,-B2)  → ~$1,607.76/month
  - B7 = =B6*B4*B5               → ~$289,396.80 total paid
  - B8 = =B7-B2                  → ~$89,396.80 total interest
  - All three target cells are EMPTY in the initial file — only task-introduced changes are scored.
"""

import os
import openpyxl

WORKDIR = '/home/user'   # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_complex_financial_068'
SHEET_NAME = 'Amortization Summary'


def normalize_formula(val):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(val, str):
        return ''
    return val.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: check the expected sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # ------------------------------------------------------------------
    # Component 1: B6 contains the PMT formula for monthly payment (0.4 pts)
    #
    # Expected formula: =PMT(B3/B5,B4*B5,-B2)
    # The exact formula uses cell references for all parameters, making it
    # dynamic. We accept any valid PMT formula that references B3/B5, B4*B5,
    # and B2 (as a negative argument), or equivalent expressions.
    #
    # Strategy:
    #   1. Verify B6 is not empty (was empty in initial file).
    #   2. Verify the formula contains PMT( and references B3, B5, B4, B2.
    # This FAILS on initial (B6 is empty) and PASSES on golden (formula present).
    # ------------------------------------------------------------------
    try:
        b6_val = ws.cell(row=6, column=2).value
        b6_norm = normalize_formula(b6_val)

        if b6_val is None or b6_val == '':
            print("FAIL: Component 1 — B6 is empty (no monthly payment formula)")
        elif not isinstance(b6_val, str) or not b6_norm.startswith('='):
            print(f"FAIL: Component 1 — B6 contains a value, not a formula: {repr(b6_val)}")
        elif 'PMT(' not in b6_norm:
            print(f"FAIL: Component 1 — B6 formula does not use PMT function: {repr(b6_val)}")
        elif all(ref in b6_norm for ref in ['B3', 'B5', 'B4', 'B2']):
            # Verify all required cell references are present in the PMT formula
            print(f"PASS: Component 1 — B6 contains PMT formula: {b6_val} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — B6 PMT formula missing expected cell refs (B3,B4,B5,B2): {repr(b6_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: B7 contains the total amount paid formula (0.3 pts)
    #
    # Expected formula: =B6*B4*B5
    # B7 should multiply monthly payment (B6) by years (B4) by payments/yr (B5).
    # This FAILS on initial (B7 is empty) and PASSES on golden (formula present).
    # ------------------------------------------------------------------
    try:
        b7_val = ws.cell(row=7, column=2).value
        b7_norm = normalize_formula(b7_val)

        if b7_val is None or b7_val == '':
            print("FAIL: Component 2 — B7 is empty (no total paid formula)")
        elif not isinstance(b7_val, str) or not b7_norm.startswith('='):
            print(f"FAIL: Component 2 — B7 contains a value, not a formula: {repr(b7_val)}")
        elif all(ref in b7_norm for ref in ['B6', 'B4', 'B5']):
            # Must reference B6 (monthly payment), B4 (years), B5 (payments/yr)
            print(f"PASS: Component 2 — B7 contains total paid formula: {b7_val} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — B7 formula does not reference B6*B4*B5: {repr(b7_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: B8 contains the total interest formula (0.3 pts)
    #
    # Expected formula: =B7-B2
    # B8 should subtract loan principal (B2) from total paid (B7).
    # This FAILS on initial (B8 is empty) and PASSES on golden (formula present).
    # ------------------------------------------------------------------
    try:
        b8_val = ws.cell(row=8, column=2).value
        b8_norm = normalize_formula(b8_val)

        if b8_val is None or b8_val == '':
            print("FAIL: Component 3 — B8 is empty (no total interest formula)")
        elif not isinstance(b8_val, str) or not b8_norm.startswith('='):
            print(f"FAIL: Component 3 — B8 contains a value, not a formula: {repr(b8_val)}")
        elif 'B7' in b8_norm and 'B2' in b8_norm and '-' in b8_norm:
            # Must subtract B2 (principal) from B7 (total paid)
            print(f"PASS: Component 3 — B8 contains total interest formula: {b8_val} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — B8 formula does not subtract B2 from B7: {repr(b8_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
