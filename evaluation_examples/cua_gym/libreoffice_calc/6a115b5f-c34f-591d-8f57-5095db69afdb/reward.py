"""
Reward Script: Verify COUNTIFS formulas in Summary sheet
Task ID: calc_gg5_017
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): B2 has COUNTIFS formula for North, Rating>=4
  Component 2 (0.25): B3 has COUNTIFS formula for South, Rating>=4
  Component 3 (0.25): B4 has COUNTIFS formula for East, Rating>=4
  Component 4 (0.25): B5 has COUNTIFS formula for West, Rating>=4
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_017'

# Expected regions in order for B2:B5
EXPECTED_REGIONS = ['North', 'South', 'East', 'West']


def is_valid_countifs(formula, region):
    """
    Check if a formula is a valid COUNTIFS that counts rows where
    Region=<region> AND Rating>=4 on the Responses sheet.

    Accepts various equivalent forms:
      =COUNTIFS(Responses.$B:$B,"North",Responses.$D:$D,">="&4)
      =COUNTIFS(Responses.B:B,"North",Responses.D:D,">=4")
      =COUNTIFS(Responses!$B:$B,"North",Responses!$D:$D,">="&4)
      etc.
    """
    if not isinstance(formula, str):
        return False, "Cell is not a formula"

    f = formula.strip()
    f_upper = f.upper()

    # Must start with =COUNTIFS(
    if not f_upper.startswith('=COUNTIFS('):
        return False, f"Expected =COUNTIFS(...), got: {f}"

    # Must reference the Responses sheet (with . or ! separator)
    if 'RESPONSES' not in f_upper:
        return False, f"Formula does not reference Responses sheet: {f}"

    # Must reference column B (Region) and column D (Rating)
    # Allow $B, B, $D, D with optional $ on row parts
    has_col_b = bool(re.search(r'RESPONSES[.!]\$?B', f_upper))
    has_col_d = bool(re.search(r'RESPONSES[.!]\$?D', f_upper))
    if not has_col_b:
        return False, f"Formula does not reference column B (Region): {f}"
    if not has_col_d:
        return False, f"Formula does not reference column D (Rating): {f}"

    # Must contain the region name (case-insensitive match in quoted string)
    if region.upper() not in f_upper:
        return False, f"Formula does not contain region '{region}': {f}"

    # Must contain a >=4 condition in some form
    # Accept: ">=4", ">="&4, ">=4", ">="&4
    has_ge4 = bool(re.search(r'>=.*4|>=.*4', f))
    if not has_ge4:
        return False, f"Formula does not contain >=4 condition: {f}"

    return True, "Valid COUNTIFS formula"


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Responses sheet must exist
    if 'Responses' not in wb.sheetnames:
        print("CRITICAL: 'Responses' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_sum = wb['Summary']

    # Also compute manual counts for cross-validation
    ws_resp = wb['Responses']
    manual_counts = {'North': 0, 'South': 0, 'East': 0, 'West': 0}
    for r in range(2, ws_resp.max_row + 1):
        region = ws_resp.cell(r, 2).value
        rating = ws_resp.cell(r, 4).value
        if region in manual_counts and rating is not None and rating >= 4:
            manual_counts[region] += 1

    # Check each region's COUNTIFS formula (B2 through B5)
    for i, region in enumerate(EXPECTED_REGIONS):
        row = i + 2  # B2, B3, B4, B5
        cell_ref = f"B{row}"
        component_weight = 0.25

        # Component: B{row} has a valid COUNTIFS formula for {region}
        try:
            cell_value = ws_sum[cell_ref].value
            valid, detail = is_valid_countifs(cell_value, region)
            if valid:
                print(f"PASS: Component {i+1} -- {cell_ref} has valid COUNTIFS for {region} ({component_weight} pts)")
                print(f"       Formula: {cell_value}")
                print(f"       Expected count: {manual_counts[region]}")
                total_score += component_weight
            else:
                print(f"FAIL: Component {i+1} -- {cell_ref} for {region}: {detail}")
        except Exception as e:
            print(f"ERROR: Component {i+1} -- checking {cell_ref}: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
