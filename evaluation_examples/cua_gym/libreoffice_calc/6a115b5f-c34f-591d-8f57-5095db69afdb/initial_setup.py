"""
Initial Setup: Create survey_results.xlsx with 800 survey responses and empty Summary sheet
Task ID: calc_gg5_017
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg5_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Sheet 1: Responses ---
    ws1 = wb.active
    ws1.title = 'Responses'

    # Headers
    headers = ['ResponseID', 'Region', 'Product', 'Rating', 'Comment']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws1.column_dimensions['A'].width = 14
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 10
    ws1.column_dimensions['E'].width = 45

    regions = ['North', 'South', 'East', 'West']
    products = [
        'UltraWidget Pro', 'SmartGadget X1', 'EcoSensor 360', 'DataPulse Mini',
        'CloudSync Hub', 'NanoTracker V2', 'PowerCore Elite', 'FlexiBoard Slim',
        'VisionLens Plus', 'AquaPure System', 'ThermoGuard Max', 'SolarBeam Lite'
    ]

    positive_comments = [
        'Excellent product quality, exceeded expectations',
        'Very satisfied with the purchase, would buy again',
        'Outstanding customer service and fast delivery',
        'Great value for money, highly recommend',
        'Works perfectly, exactly as described',
        'Impressive build quality and durability',
        'Love the design and functionality',
        'Best purchase I have made this year',
        'Superb performance in all conditions',
        'Top notch quality and reliable operation',
    ]
    neutral_comments = [
        'Decent product, meets basic requirements',
        'Average experience, nothing special',
        'Product is okay but could be improved',
        'Acceptable quality for the price range',
        'Standard performance, no complaints',
        'Fair product, adequate for daily use',
        'Meets expectations but does not exceed them',
        'Reasonable quality, room for improvement',
    ]
    negative_comments = [
        'Disappointed with the overall quality',
        'Did not meet my expectations at all',
        'Product arrived damaged, poor packaging',
        'Not worth the price, would not recommend',
        'Functionality is limited compared to description',
        'Poor durability, broke within a week',
    ]

    # Generate 800 rows of realistic survey data
    for i in range(1, 801):
        row = i + 1
        response_id = f'RESP-{10000 + i}'
        region = random.choice(regions)
        product = random.choice(products)
        rating = random.choices([1, 2, 3, 4, 5], weights=[5, 10, 20, 35, 30])[0]

        if rating >= 4:
            comment = random.choice(positive_comments)
        elif rating == 3:
            comment = random.choice(neutral_comments)
        else:
            comment = random.choice(negative_comments)

        ws1.cell(row=row, column=1, value=response_id)
        ws1.cell(row=row, column=2, value=region)
        ws1.cell(row=row, column=3, value=product)
        ws1.cell(row=row, column=4, value=rating)
        ws1.cell(row=row, column=5, value=comment)

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')

    # Headers
    ws2.cell(row=1, column=1, value='Region')
    ws2.cell(row=1, column=1).font = Font(bold=True, size=11)
    ws2.cell(row=1, column=1).fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    ws2.cell(row=1, column=2, value='High Ratings (4+)')
    ws2.cell(row=1, column=2).font = Font(bold=True, size=11)
    ws2.cell(row=1, column=2).fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")

    # Region names in A2:A5
    ws2.cell(row=2, column=1, value='North')
    ws2.cell(row=3, column=1, value='South')
    ws2.cell(row=4, column=1, value='East')
    ws2.cell(row=5, column=1, value='West')

    # B2:B5 left EMPTY - agent must fill with COUNTIFS formulas
    # Do NOT put any values or formulas here

    ws2.column_dimensions['A'].width = 14
    ws2.column_dimensions['B'].width = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
