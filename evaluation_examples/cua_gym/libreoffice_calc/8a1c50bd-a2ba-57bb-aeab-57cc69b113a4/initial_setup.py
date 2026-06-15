"""
Initial Setup: Build a complete reference manager for 12 papers
Task ID: osworld_multi_apps_web_references_011
Domain: libreoffice_calc (multi-app: Calc, Writer, OS)

Creates: /home/user/Desktop/papers_incomplete.ods
  - 12 classic NLP/CV/RL papers (2012-2023)
  - Columns: Title, Authors, Year, DOI, OA_PDF_URL, Citation_Count, Venue, BibTeX_Key
  - DOI, OA_PDF_URL, Citation_Count, Venue are all EMPTY (task is to fill these)
  - BibTeX_Key is pre-filled

Then opens the file in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
DOCUMENTS = '/home/user/Documents'
OUTPUT = f'{DESKTOP}/papers_incomplete.ods'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop and Documents directories exist
    os.makedirs(DESKTOP, exist_ok=True)
    os.makedirs(DOCUMENTS, exist_ok=True)

    # Try to use openpyxl first (for xlsx), then write ODS using odfpy or subprocess
    # Since the task specifies .ods format, we'll create it as .xlsx first then convert,
    # OR use the python-odf library if available.
    # We use openpyxl to write a temporary xlsx, then convert to ods using LibreOffice CLI.

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        TEMP_XLSX = f'{WORKDIR}/papers_incomplete_temp.xlsx'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Papers"

        # Headers
        headers = ['Title', 'Authors', 'Year', 'DOI', 'OA_PDF_URL', 'Citation_Count', 'Venue', 'BibTeX_Key']
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, name='Calibri', size=11)
            cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
            cell.font = Font(bold=True, name='Calibri', size=11, color="FFFFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # 12 classic NLP/CV/RL papers, 2012-2023
        # Columns: Title, Authors, Year, DOI(empty), OA_PDF_URL(empty), Citation_Count(empty), Venue(empty), BibTeX_Key(pre-filled)
        papers = [
            # NLP papers
            (
                "Attention Is All You Need",
                "Vaswani, A., Shazeer, N., Parmar, N., Uszkoreit, J., Jones, L., Gomez, A. N., Kaiser, L., & Polosukhin, I.",
                2017,
                "", "", "", "",
                "vaswani2017attention"
            ),
            (
                "BERT: Pre-training of Deep Bidirectional Transformers for Language Understanding",
                "Devlin, J., Chang, M. W., Lee, K., & Toutanova, K.",
                2019,
                "", "", "", "",
                "devlin2019bert"
            ),
            (
                "Language Models are Few-Shot Learners",
                "Brown, T. B., Mann, B., Ryder, N., Subbiah, M., Kaplan, J., Dhariwal, P., Neelakantan, A., Shyam, P., Sastry, G., Askell, A., et al.",
                2020,
                "", "", "", "",
                "brown2020gpt3"
            ),
            (
                "Improving Language Understanding by Generative Pre-Training",
                "Radford, A., Narasimhan, K., Salimans, T., & Sutskever, I.",
                2018,
                "", "", "", "",
                "radford2018gpt"
            ),
            # CV papers
            (
                "ImageNet Classification with Deep Convolutional Neural Networks",
                "Krizhevsky, A., Sutskever, I., & Hinton, G. E.",
                2012,
                "", "", "", "",
                "krizhevsky2012imagenet"
            ),
            (
                "Deep Residual Learning for Image Recognition",
                "He, K., Zhang, X., Ren, S., & Sun, J.",
                2016,
                "", "", "", "",
                "he2016resnet"
            ),
            (
                "An Image is Worth 16x16 Words: Transformers for Image Recognition at Scale",
                "Dosovitskiy, A., Beyer, L., Kolesnikov, A., Weissenborn, D., Zhai, X., Unterthiner, T., Dehghani, M., Minderer, M., Heigold, G., Gelly, S., Uszkoreit, J., & Houlsby, N.",
                2021,
                "", "", "", "",
                "dosovitskiy2021vit"
            ),
            (
                "You Only Look Once: Unified, Real-Time Object Detection",
                "Redmon, J., Divvala, S., Girshick, R., & Farhadi, A.",
                2016,
                "", "", "", "",
                "redmon2016yolo"
            ),
            # RL papers
            (
                "Human-level control through deep reinforcement learning",
                "Mnih, V., Kavukcuoglu, K., Silver, D., Rusu, A. A., Veness, J., Bellemare, M. G., Graves, A., Riedmiller, M., Fidjeland, A. K., Ostrovski, G., et al.",
                2015,
                "", "", "", "",
                "mnih2015dqn"
            ),
            (
                "Mastering the game of Go with deep neural networks and tree search",
                "Silver, D., Huang, A., Maddison, C. J., Guez, A., Sifre, L., van den Driessche, G., Schrittwieser, J., Antonoglou, I., Panneershelvam, V., Lanctot, M., et al.",
                2016,
                "", "", "", "",
                "silver2016alphago"
            ),
            (
                "Proximal Policy Optimization Algorithms",
                "Schulman, J., Wolski, F., Dhariwal, P., Radford, A., & Klimov, O.",
                2017,
                "", "", "", "",
                "schulman2017ppo"
            ),
            (
                "Training language models to follow instructions with human feedback",
                "Ouyang, L., Wu, J., Jiang, X., Almeida, D., Wainwright, C., Mishkin, P., Zhang, C., Agarwal, S., Slama, K., Ray, A., et al.",
                2022,
                "", "", "", "",
                "ouyang2022instructgpt"
            ),
        ]

        for r, paper in enumerate(papers, 2):
            for c, val in enumerate(paper, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Column widths for readability
        ws.column_dimensions['A'].width = 55  # Title
        ws.column_dimensions['B'].width = 50  # Authors
        ws.column_dimensions['C'].width = 8   # Year
        ws.column_dimensions['D'].width = 30  # DOI (empty)
        ws.column_dimensions['E'].width = 40  # OA_PDF_URL (empty)
        ws.column_dimensions['F'].width = 15  # Citation_Count (empty)
        ws.column_dimensions['G'].width = 30  # Venue (empty)
        ws.column_dimensions['H'].width = 25  # BibTeX_Key

        # Freeze header row
        ws.freeze_panes = "A2"

        wb.save(TEMP_XLSX)
        print(f'Temporary XLSX created: {TEMP_XLSX}')

        # Convert xlsx -> ods using LibreOffice headless
        env = os.environ.copy()
        env["DISPLAY"] = ":0"
        result = subprocess.run(
            ['libreoffice', '--headless', '--convert-to', 'ods', '--outdir', DESKTOP, TEMP_XLSX],
            capture_output=True,
            text=True,
            env=env,
            timeout=60
        )
        print(f'Conversion stdout: {result.stdout}')
        print(f'Conversion stderr: {result.stderr}')

        # Clean up temp file
        if os.path.exists(TEMP_XLSX):
            os.remove(TEMP_XLSX)

        if os.path.exists(OUTPUT):
            print(f'ODS file created at: {OUTPUT}')
        else:
            # Fallback: try to write ODS directly using odfpy
            print('ODS conversion failed, trying odfpy...')
            create_ods_with_odfpy()

    except ImportError:
        print('openpyxl not available, trying odfpy directly...')
        create_ods_with_odfpy()

    # GUI-ready startup: open the papers file in LibreOffice Calc
    time.sleep(2)
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print(f'GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


def create_ods_with_odfpy():
    """Fallback: create ODS directly using odfpy library."""
    try:
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TextProperties, TableCellProperties

        doc = OpenDocumentSpreadsheet()
        table = Table(name="Papers")

        headers = ['Title', 'Authors', 'Year', 'DOI', 'OA_PDF_URL', 'Citation_Count', 'Venue', 'BibTeX_Key']
        papers = [
            ("Attention Is All You Need",
             "Vaswani, A., Shazeer, N., Parmar, N., Uszkoreit, J., Jones, L., Gomez, A. N., Kaiser, L., & Polosukhin, I.",
             2017, "", "", "", "", "vaswani2017attention"),
            ("BERT: Pre-training of Deep Bidirectional Transformers for Language Understanding",
             "Devlin, J., Chang, M. W., Lee, K., & Toutanova, K.",
             2019, "", "", "", "", "devlin2019bert"),
            ("Language Models are Few-Shot Learners",
             "Brown, T. B., Mann, B., Ryder, N., Subbiah, M., Kaplan, J., Dhariwal, P., et al.",
             2020, "", "", "", "", "brown2020gpt3"),
            ("Improving Language Understanding by Generative Pre-Training",
             "Radford, A., Narasimhan, K., Salimans, T., & Sutskever, I.",
             2018, "", "", "", "", "radford2018gpt"),
            ("ImageNet Classification with Deep Convolutional Neural Networks",
             "Krizhevsky, A., Sutskever, I., & Hinton, G. E.",
             2012, "", "", "", "", "krizhevsky2012imagenet"),
            ("Deep Residual Learning for Image Recognition",
             "He, K., Zhang, X., Ren, S., & Sun, J.",
             2016, "", "", "", "", "he2016resnet"),
            ("An Image is Worth 16x16 Words: Transformers for Image Recognition at Scale",
             "Dosovitskiy, A., Beyer, L., Kolesnikov, A., Weissenborn, D., Zhai, X., Unterthiner, T., et al.",
             2021, "", "", "", "", "dosovitskiy2021vit"),
            ("You Only Look Once: Unified, Real-Time Object Detection",
             "Redmon, J., Divvala, S., Girshick, R., & Farhadi, A.",
             2016, "", "", "", "", "redmon2016yolo"),
            ("Human-level control through deep reinforcement learning",
             "Mnih, V., Kavukcuoglu, K., Silver, D., Rusu, A. A., Veness, J., Bellemare, M. G., et al.",
             2015, "", "", "", "", "mnih2015dqn"),
            ("Mastering the game of Go with deep neural networks and tree search",
             "Silver, D., Huang, A., Maddison, C. J., Guez, A., Sifre, L., et al.",
             2016, "", "", "", "", "silver2016alphago"),
            ("Proximal Policy Optimization Algorithms",
             "Schulman, J., Wolski, F., Dhariwal, P., Radford, A., & Klimov, O.",
             2017, "", "", "", "", "schulman2017ppo"),
            ("Training language models to follow instructions with human feedback",
             "Ouyang, L., Wu, J., Jiang, X., Almeida, D., Wainwright, C., Mishkin, P., et al.",
             2022, "", "", "", "", "ouyang2022instructgpt"),
        ]

        # Header row
        hr = TableRow()
        for h in headers:
            tc = TableCell(valuetype="string")
            tc.addElement(P(text=str(h)))
            hr.addElement(tc)
        table.addElement(hr)

        # Data rows
        for paper in papers:
            row = TableRow()
            for val in paper:
                if isinstance(val, int):
                    tc = TableCell(valuetype="float", value=str(val))
                    tc.addElement(P(text=str(val)))
                else:
                    tc = TableCell(valuetype="string")
                    tc.addElement(P(text=str(val)))
                row.addElement(tc)
            table.addElement(row)

        doc.spreadsheet.addElement(table)
        doc.save(OUTPUT)
        print(f'ODS file created (odfpy): {OUTPUT}')

    except ImportError as e:
        print(f'odfpy not available: {e}')
        print('Creating CSV fallback...')
        # Last resort: write as CSV (agent can still work with it)
        import csv
        csv_path = OUTPUT.replace('.ods', '.csv')
        papers_data = [
            ['Title', 'Authors', 'Year', 'DOI', 'OA_PDF_URL', 'Citation_Count', 'Venue', 'BibTeX_Key'],
            ['Attention Is All You Need', 'Vaswani, A. et al.', 2017, '', '', '', '', 'vaswani2017attention'],
            ['BERT: Pre-training of Deep Bidirectional Transformers', 'Devlin, J. et al.', 2019, '', '', '', '', 'devlin2019bert'],
        ]
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(papers_data)
        print(f'CSV fallback created: {csv_path}')


create_initial()
