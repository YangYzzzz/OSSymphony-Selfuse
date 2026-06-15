"""
Initial Setup: Small Business Cash Flow Forecaster
Task ID: calc_gen_smallbiz_066
Domain: libreoffice_calc

Creates a workbook with two sheets: 'Assumptions' and 'CashFlow'
- Assumptions sheet: input parameters (starting cash, revenue, growth rate,
  fixed expenses, variable expense rate)
- CashFlow sheet: month labels and row labels only — NO formulas yet
  (the agent's task is to add all formulas and conditional formatting)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_smallbiz_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ── Sheet 1: Assumptions ────────────────────────────────────────────────
    ws_a = wb.active
    ws_a.title = 'Assumptions'

    # Title
    ws_a['A1'] = 'Small Business Cash Flow Forecaster — Assumptions'
    ws_a['A1'].font = Font(bold=True, size=13)
    ws_a.merge_cells('A1:C1')

    # Input fields
    ws_a['A2'] = 'Starting Cash Balance'
    ws_a['B2'] = 50000
    ws_a['B2'].number_format = '$#,##0.00'

    ws_a['A3'] = 'Month 1 Revenue'
    ws_a['B3'] = 35000
    ws_a['B3'].number_format = '$#,##0.00'

    ws_a['A4'] = 'Monthly Revenue Growth Rate'
    ws_a['B4'] = 0.05
    ws_a['B4'].number_format = '0.00%'

    # Header for fixed expenses
    ws_a['A5'] = 'Rent / Lease'
    ws_a['B5'] = 4500
    ws_a['B5'].number_format = '$#,##0.00'

    ws_a['A6'] = 'Salaries & Wages'
    ws_a['B6'] = 12000
    ws_a['B6'].number_format = '$#,##0.00'

    ws_a['A7'] = 'Utilities'
    ws_a['B7'] = 850
    ws_a['B7'].number_format = '$#,##0.00'

    ws_a['A8'] = 'Insurance'
    ws_a['B8'] = 600
    ws_a['B8'].number_format = '$#,##0.00'

    ws_a['A9'] = 'Loan Repayment'
    ws_a['B9'] = 1800
    ws_a['B9'].number_format = '$#,##0.00'

    ws_a['A10'] = 'Marketing & Advertising'
    ws_a['B10'] = 1200
    ws_a['B10'].number_format = '$#,##0.00'

    ws_a['A11'] = 'Variable Expense Rate (% of Revenue)'
    ws_a['B11'] = 0.30
    ws_a['B11'].number_format = '0.00%'

    # Labels / notes column
    ws_a['C2'] = 'Cash on hand at start of Jan'
    ws_a['C3'] = 'Base revenue for January'
    ws_a['C4'] = '← Sensitivity input: change this to update entire forecast'
    ws_a['C4'].font = Font(italic=True, color='FF0070C0')
    ws_a['C5'] = 'Fixed cost'
    ws_a['C6'] = 'Fixed cost'
    ws_a['C7'] = 'Fixed cost'
    ws_a['C8'] = 'Fixed cost'
    ws_a['C9'] = 'Fixed cost'
    ws_a['C10'] = 'Fixed cost'
    ws_a['C11'] = 'Scales with monthly revenue'

    # Style column A labels bold
    for row in range(2, 12):
        ws_a.cell(row=row, column=1).font = Font(bold=True)

    # Column widths
    ws_a.column_dimensions['A'].width = 34
    ws_a.column_dimensions['B'].width = 16
    ws_a.column_dimensions['C'].width = 46

    # ── Sheet 2: CashFlow ───────────────────────────────────────────────────
    ws_c = wb.create_sheet('CashFlow')

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

    # Row 1: Month labels
    ws_c['A1'] = 'Month'
    for i, m in enumerate(months):
        ws_c.cell(row=1, column=i + 2, value=m)

    # Row 2: Revenue (no formulas — agent must add them)
    ws_c['A2'] = 'Revenue'

    # Row 3: Fixed Expenses (no formulas)
    ws_c['A3'] = 'Fixed Expenses'

    # Row 4: Variable Expenses (no formulas)
    ws_c['A4'] = 'Variable Expenses'

    # Row 5: Net Cash Flow (no formulas)
    ws_c['A5'] = 'Net Cash Flow'

    # Row 6: Cumulative Cash Position (no formulas, no conditional formatting)
    ws_c['A6'] = 'Cumulative Cash Position'

    # Style header row bold
    for col in range(1, 14):
        ws_c.cell(row=1, column=col).font = Font(bold=True)

    # Style column A (row labels) bold
    for row in range(1, 7):
        ws_c.cell(row=row, column=1).font = Font(bold=True)

    # Column widths
    ws_c.column_dimensions['A'].width = 26
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws_c.column_dimensions[col_letter].width = 12

    # Number format placeholder on data rows (columns B:M, rows 2-6)
    for row in range(2, 7):
        for col in range(2, 14):
            ws_c.cell(row=row, column=col).number_format = '$#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Assumptions, CashFlow')
    print('CashFlow has month labels and row labels — NO formulas or conditional formatting yet')


create_initial()
