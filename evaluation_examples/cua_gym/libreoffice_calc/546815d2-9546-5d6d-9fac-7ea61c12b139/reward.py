"""
Reward Script: Small Business Cash Flow Forecaster
Task ID: calc_gen_smallbiz_066
Domain: libreoffice_calc
Scoring:
  - Component 1: Revenue formulas in CashFlow B2:M2 linked to Assumptions (0.25 pts)
  - Component 2: Fixed Expenses formulas in CashFlow B3:M3 summing Assumptions fixed costs (0.20 pts)
  - Component 3: Variable Expenses formulas in CashFlow B4:M4 as revenue * variable rate (0.15 pts)
  - Component 4: Net Cash Flow formulas in CashFlow B5:M5 = revenue - fixed - variable (0.15 pts)
  - Component 5: Cumulative Cash formulas in CashFlow B6:M6 = cumulative sum from starting balance (0.15 pts)
  - Component 6: Conditional formatting on row 6: red fill when value < 0 (0.10 pts)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_smallbiz_066'


def normalize_formula(f):
    """Normalize formula for comparison: uppercase, remove spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Gate: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: required sheets must exist
    if 'CashFlow' not in wb.sheetnames or 'Assumptions' not in wb.sheetnames:
        print(f"FAIL: Required sheets missing. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_cf = wb['CashFlow']
    ws_as = wb['Assumptions']

    # -------------------------------------------------------------------------
    # Component 1: Revenue row (B2:M2) — 0.25 points
    #   B2 should reference Assumptions!B3 (or Assumptions!$B$3)
    #   C2:M2 should each be prior_month * (1 + Assumptions!$B$4)
    # -------------------------------------------------------------------------
    try:
        revenue_cols = [chr(ord('B') + i) for i in range(12)]  # B..M

        # Check B2: links to Assumptions!B3
        b2_val = ws_cf['B2'].value
        b2_ok = False
        if isinstance(b2_val, str):
            norm = normalize_formula(b2_val)
            # Accept various forms: =ASSUMPTIONS!B3, =ASSUMPTIONS!$B$3, etc.
            if re.search(r'ASSUMPTIONS.*B.*3', norm):
                b2_ok = True

        # Check C2:M2: each = prior*(1+Assumptions!$B$4)
        growth_formulas_ok = 0
        growth_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        for i, col in enumerate(growth_cols):
            prev_col = chr(ord(col) - 1)
            cell_val = ws_cf[f'{col}2'].value
            if isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Should reference prior column row2 and Assumptions B4 (possibly with $ signs)
                has_prev_col = f'{prev_col}2' in norm
                # Match $B$4 or B4
                has_growth_rate = re.search(r'\$?B\$?4', norm) is not None
                if has_prev_col and has_growth_rate and 'ASSUMPTIONS' in norm:
                    growth_formulas_ok += 1

        if b2_ok and growth_formulas_ok >= 9:
            print(f"PASS: Component 1 — Revenue formulas B2:M2 correctly linked to Assumptions "
                  f"(B2 links Assumptions!B3, {growth_formulas_ok}/11 growth formulas correct) (0.25 pts)")
            total_score += 0.25
        elif b2_ok and growth_formulas_ok >= 5:
            print(f"PASS (partial): Component 1 — B2 links Assumptions!B3, "
                  f"but only {growth_formulas_ok}/11 growth formulas correct (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Revenue formulas incorrect. "
                  f"B2 ok={b2_ok} (value={repr(b2_val)}), growth formulas ok={growth_formulas_ok}/11")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Fixed Expenses row (B3:M3) — 0.20 points
    #   All 12 months should use SUM(Assumptions!$B$5:$B$10)
    # -------------------------------------------------------------------------
    try:
        fixed_ok = 0
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            cell_val = ws_cf[f'{col}3'].value
            if isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Must reference Assumptions and SUM B5:B10 (with or without $)
                has_b5 = re.search(r'\$?B\$?5', norm) is not None
                has_b10 = re.search(r'\$?B\$?10', norm) is not None
                if 'SUM' in norm and 'ASSUMPTIONS' in norm and has_b5 and has_b10:
                    fixed_ok += 1

        if fixed_ok >= 10:
            print(f"PASS: Component 2 — Fixed Expenses formulas {fixed_ok}/12 reference "
                  f"SUM(Assumptions!$B$5:$B$10) (0.20 pts)")
            total_score += 0.20
        elif fixed_ok >= 5:
            print(f"PASS (partial): Component 2 — Fixed Expenses: only {fixed_ok}/12 correct (0.10 pts)")
            total_score += 0.10
        else:
            # Sample B3 for debug
            sample = ws_cf['B3'].value
            print(f"FAIL: Component 2 — Fixed Expenses formulas incorrect. "
                  f"Only {fixed_ok}/12 correct. B3={repr(sample)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Variable Expenses row (B4:M4) — 0.15 points
    #   Each cell = revenue_cell * Assumptions!$B$11
    # -------------------------------------------------------------------------
    try:
        var_ok = 0
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            cell_val = ws_cf[f'{col}4'].value
            if isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Must reference same column row2 and Assumptions B11 (with or without $)
                has_rev_ref = f'{col}2' in norm
                has_var_rate = re.search(r'\$?B\$?11', norm) is not None
                if has_rev_ref and 'ASSUMPTIONS' in norm and has_var_rate:
                    var_ok += 1

        if var_ok >= 10:
            print(f"PASS: Component 3 — Variable Expenses formulas {var_ok}/12 "
                  f"reference revenue*Assumptions!$B$11 (0.15 pts)")
            total_score += 0.15
        elif var_ok >= 5:
            print(f"PASS (partial): Component 3 — Variable Expenses: only {var_ok}/12 correct (0.07 pts)")
            total_score += 0.07
        else:
            sample = ws_cf['B4'].value
            print(f"FAIL: Component 3 — Variable Expenses formulas incorrect. "
                  f"Only {var_ok}/12 correct. B4={repr(sample)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Net Cash Flow row (B5:M5) — 0.15 points
    #   Each cell = revenue - fixed - variable (e.g., =B2-B3-B4)
    # -------------------------------------------------------------------------
    try:
        net_ok = 0
        for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
            cell_val = ws_cf[f'{col}5'].value
            if isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Must reference same column rows 2, 3, 4 with subtraction
                if f'{col}2' in norm and f'{col}3' in norm and f'{col}4' in norm and '-' in norm:
                    net_ok += 1

        if net_ok >= 10:
            print(f"PASS: Component 4 — Net Cash Flow formulas {net_ok}/12 "
                  f"subtract fixed and variable from revenue (0.15 pts)")
            total_score += 0.15
        elif net_ok >= 5:
            print(f"PASS (partial): Component 4 — Net Cash Flow: only {net_ok}/12 correct (0.07 pts)")
            total_score += 0.07
        else:
            sample = ws_cf['B5'].value
            print(f"FAIL: Component 4 — Net Cash Flow formulas incorrect. "
                  f"Only {net_ok}/12 correct. B5={repr(sample)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Cumulative Cash Position row (B6:M6) — 0.15 points
    #   B6 = Assumptions!$B$2 + B5 (starting balance + first net cash flow)
    #   C6:M6 = prior cumulative + current net cash flow
    # -------------------------------------------------------------------------
    try:
        b6_val = ws_cf['B6'].value
        b6_ok = False
        if isinstance(b6_val, str):
            norm = normalize_formula(b6_val)
            # B6 should reference Assumptions B2 (starting balance, possibly with $) + B5
            has_b2 = re.search(r'\$?B\$?2', norm) is not None
            has_b5 = 'B5' in norm
            if 'ASSUMPTIONS' in norm and has_b2 and has_b5:
                b6_ok = True

        cum_ok = 0
        cum_cols = ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
        for i, col in enumerate(cum_cols):
            prev_col = chr(ord(col) - 1)
            cell_val = ws_cf[f'{col}6'].value
            if isinstance(cell_val, str):
                norm = normalize_formula(cell_val)
                # Should reference prior col row 6 and same col row 5
                if f'{prev_col}6' in norm and f'{col}5' in norm:
                    cum_ok += 1

        if b6_ok and cum_ok >= 9:
            print(f"PASS: Component 5 — Cumulative Cash formulas correct "
                  f"(B6 refs Assumptions!B2, {cum_ok}/11 subsequent months correct) (0.15 pts)")
            total_score += 0.15
        elif b6_ok and cum_ok >= 5:
            print(f"PASS (partial): Component 5 — B6 correct but only {cum_ok}/11 "
                  f"subsequent months correct (0.07 pts)")
            total_score += 0.07
        elif not b6_ok and cum_ok >= 9:
            print(f"PASS (partial): Component 5 — Subsequent months correct ({cum_ok}/11) "
                  f"but B6 incorrect (value={repr(b6_val)}) (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 5 — Cumulative Cash formulas incorrect. "
                  f"B6 ok={b6_ok} (value={repr(b6_val)}), subsequent months ok={cum_ok}/11")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------------
    # Component 6: Conditional formatting on B6:M6 — 0.10 points
    #   Red fill when value < 0 (cellIs lessThan 0)
    # -------------------------------------------------------------------------
    try:
        cf_found = False
        for cf_range, cf_list in ws_cf.conditional_formatting._cf_rules.items():
            range_str = str(cf_range)
            # Check range covers row 6
            if '6' in range_str:
                for rule in cf_list:
                    # Should be a cellIs rule with lessThan 0 and red fill
                    is_less_than_zero = (
                        rule.type == 'cellIs' and
                        rule.operator == 'lessThan' and
                        rule.formula and
                        '0' in str(rule.formula)
                    )
                    has_red_fill = False
                    if rule.dxf and rule.dxf.fill:
                        try:
                            fill_color = rule.dxf.fill.fgColor.rgb
                            # Accept any red-ish color (FF in ARGB with high R value)
                            if 'FF' in fill_color.upper() and fill_color.upper().startswith('FF'):
                                # FFFF0000 is pure red
                                has_red_fill = fill_color.upper() == 'FFFF0000'
                            # Also accept just presence of a fill rule
                            if rule.dxf.fill.patternType == 'solid':
                                has_red_fill = True
                        except Exception:
                            if rule.dxf.fill.patternType:
                                has_red_fill = True

                    if is_less_than_zero and has_red_fill:
                        cf_found = True
                        break
            if cf_found:
                break

        if cf_found:
            print("PASS: Component 6 — Conditional formatting on row 6: "
                  "red fill when < 0 (0.10 pts)")
            total_score += 0.10
        else:
            # Check if any CF exists on row 6 at all (partial)
            any_cf_on_row6 = False
            for cf_range, cf_list in ws_cf.conditional_formatting._cf_rules.items():
                if '6' in str(cf_range) and cf_list:
                    any_cf_on_row6 = True
                    break
            if any_cf_on_row6:
                print("PASS (partial): Component 6 — CF exists on row 6 but not exactly "
                      "red-fill-when-negative (0.05 pts)")
                total_score += 0.05
            else:
                print("FAIL: Component 6 — No conditional formatting found on Cumulative Cash row 6")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
