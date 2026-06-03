"""
Reward Script: Verify Pivot Table creation from HR Data
Task ID: calc_ggf_028
Domain: libreoffice_calc
Scoring:
  Component 1 (0.10): A new sheet exists beyond 'HR Data' with pivot table data
  Component 2 (0.20): Count of Employee ID section has correct structure (Department rows, F/M columns)
  Component 3 (0.25): Count values are correct
  Component 4 (0.20): Average of Salary section has correct structure
  Component 5 (0.25): Average salary values are correct (with tolerance)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_028'

# Expected ground truth values from the golden pivot table
DEPARTMENTS = ['Engineering', 'Finance', 'HR', 'Marketing', 'Operations', 'Sales']

EXPECTED_COUNTS = {
    'Engineering': {'F': 25, 'M': 25, 'Grand Total': 50},
    'Finance':     {'F': 25, 'M': 25, 'Grand Total': 50},
    'HR':          {'F': 25, 'M': 25, 'Grand Total': 50},
    'Marketing':   {'F': 25, 'M': 25, 'Grand Total': 50},
    'Operations':  {'F': 25, 'M': 25, 'Grand Total': 50},
    'Sales':       {'F': 25, 'M': 25, 'Grand Total': 50},
    'Grand Total': {'F': 150, 'M': 150, 'Grand Total': 300},
}

EXPECTED_AVERAGES = {
    'Engineering': {'F': 93855.47, 'M': 93740.56, 'Grand Total': 93798.02},
    'Finance':     {'F': 88070.79, 'M': 91043.88, 'Grand Total': 89557.34},
    'HR':          {'F': 66151.71, 'M': 63950.69, 'Grand Total': 65051.20},
    'Marketing':   {'F': 74206.48, 'M': 76923.32, 'Grand Total': 75564.90},
    'Operations':  {'F': 66824.25, 'M': 62047.45, 'Grand Total': 64435.85},
    'Sales':       {'F': 76147.52, 'M': 81945.93, 'Grand Total': 79046.72},
    'Grand Total': {'F': 77542.70, 'M': 78275.30, 'Grand Total': 77909.00},
}


def find_pivot_sheet(wb):
    """Find the sheet that contains pivot table data (any sheet other than 'HR Data')."""
    for name in wb.sheetnames:
        if name.lower() != 'hr data':
            return wb[name]
    return None


def find_section(ws, keyword):
    """
    Find a section in the pivot table sheet by scanning for a row containing the keyword.
    Returns (header_row, col_map, data_start_row) where:
      - header_row is the row with Department / F / M / Grand Total headers
      - col_map maps header labels to column indices
      - data_start_row is the first data row after the header
    """
    for row_idx in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and keyword.lower() in str(cell_val).lower():
            # The next row should be the header row (Department, F, M, Grand Total)
            header_row = row_idx + 1
            col_map = {}
            for col_idx in range(1, ws.max_column + 1):
                hdr = ws.cell(row=header_row, column=col_idx).value
                if hdr is not None:
                    col_map[str(hdr).strip()] = col_idx
            data_start_row = header_row + 1
            return header_row, col_map, data_start_row
    return None, None, None


def read_section_data(ws, col_map, data_start_row, num_rows=7):
    """
    Read data rows from a section.
    Returns dict: {department_name: {col_header: value, ...}, ...}
    """
    dept_col = None
    # Find the department/row-label column (first column in col_map)
    for key, idx in col_map.items():
        if key.lower() in ('department', 'dept'):
            dept_col = idx
            break
    if dept_col is None:
        # Use column 1 as default
        dept_col = min(col_map.values())

    result = {}
    for r in range(data_start_row, data_start_row + num_rows + 5):
        dept_val = ws.cell(row=r, column=dept_col).value
        if dept_val is None:
            break
        dept_name = str(dept_val).strip()
        row_data = {}
        for hdr, cidx in col_map.items():
            if cidx == dept_col:
                continue
            row_data[str(hdr).strip()] = ws.cell(row=r, column=cidx).value
        result[dept_name] = row_data
    return result


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A new sheet exists beyond 'HR Data' (0.10 points)
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Found pivot sheet '{pivot_ws.title}' (0.10 pts)")
            total_score += 0.10
        else:
            print("FAIL: Component 1 — No sheet found beyond 'HR Data'")
            # No pivot sheet means nothing else to check
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Count of Employee ID section structure (0.20 points)
    try:
        count_hdr_row, count_col_map, count_data_start = find_section(pivot_ws, 'Count')
        if count_hdr_row is not None and count_col_map:
            # Check we have F and M columns
            has_f = any('F' == k.strip() for k in count_col_map)
            has_m = any('M' == k.strip() for k in count_col_map)
            has_dept = any(k.lower() in ('department', 'dept') for k in count_col_map)
            if has_f and has_m:
                print(f"PASS: Component 2 — Count section found with F/M columns at row {count_hdr_row} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — Count section found but missing F/M columns. Headers: {count_col_map}")
        else:
            print("FAIL: Component 2 — Could not find 'Count of Employee ID' section")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Count values are correct (0.25 points)
    try:
        if count_hdr_row is not None and count_col_map:
            count_data = read_section_data(pivot_ws, count_col_map, count_data_start)
            correct_counts = 0
            total_checks = 0
            for dept, expected_vals in EXPECTED_COUNTS.items():
                if dept in count_data:
                    for col_key, expected_val in expected_vals.items():
                        actual = count_data[dept].get(col_key)
                        if actual is not None:
                            try:
                                if abs(float(actual) - float(expected_val)) < 1.0:
                                    correct_counts += 1
                            except (ValueError, TypeError):
                                pass
                        total_checks += 1
                else:
                    total_checks += len(expected_vals)

            if total_checks > 0:
                ratio = correct_counts / total_checks
                pts = round(0.25 * ratio, 4)
                total_score += pts
                print(f"PASS: Component 3 — Count values {correct_counts}/{total_checks} correct ({pts} pts)")
            else:
                print("FAIL: Component 3 — No count data found to verify")
        else:
            print("FAIL: Component 3 — Skipped (no count section)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Average of Salary section structure (0.20 points)
    try:
        avg_hdr_row, avg_col_map, avg_data_start = find_section(pivot_ws, 'Average')
        if avg_hdr_row is not None and avg_col_map:
            has_f = any('F' == k.strip() for k in avg_col_map)
            has_m = any('M' == k.strip() for k in avg_col_map)
            if has_f and has_m:
                print(f"PASS: Component 4 — Average section found with F/M columns at row {avg_hdr_row} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 4 — Average section found but missing F/M columns. Headers: {avg_col_map}")
        else:
            print("FAIL: Component 4 — Could not find 'Average of Salary' section")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Average salary values are correct with tolerance (0.25 points)
    try:
        if avg_hdr_row is not None and avg_col_map:
            avg_data = read_section_data(pivot_ws, avg_col_map, avg_data_start)
            correct_avgs = 0
            total_checks = 0
            for dept, expected_vals in EXPECTED_AVERAGES.items():
                if dept in avg_data:
                    for col_key, expected_val in expected_vals.items():
                        actual = avg_data[dept].get(col_key)
                        if actual is not None:
                            try:
                                # Allow tolerance of 5.0 for rounding differences
                                if abs(float(actual) - float(expected_val)) < 5.0:
                                    correct_avgs += 1
                            except (ValueError, TypeError):
                                pass
                        total_checks += 1
                else:
                    total_checks += len(expected_vals)

            if total_checks > 0:
                ratio = correct_avgs / total_checks
                pts = round(0.25 * ratio, 4)
                total_score += pts
                print(f"PASS: Component 5 — Average values {correct_avgs}/{total_checks} correct ({pts} pts)")
            else:
                print("FAIL: Component 5 — No average data found to verify")
        else:
            print("FAIL: Component 5 — Skipped (no average section)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
