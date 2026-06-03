"""
Initial Setup: Create HR Data spreadsheet with 300 employee records
Task ID: calc_ggf_028
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HR Data'

    # --- Headers ---
    headers = ['Employee ID', 'Department', 'Gender', 'Salary', 'Years']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Data Generation ---
    departments = ['Engineering', 'Marketing', 'Finance', 'HR', 'Operations', 'Sales']
    genders = ['M', 'F']

    # Salary ranges by department (base_min, base_max)
    salary_ranges = {
        'Engineering': (65000, 130000),
        'Marketing': (50000, 95000),
        'Finance': (60000, 120000),
        'HR': (48000, 90000),
        'Operations': (45000, 85000),
        'Sales': (52000, 110000),
    }

    # Generate 300 records ensuring both genders in every department
    # First create balanced pairs: 25 M + 25 F per department = 300 total
    assignments = []
    for dept in departments:
        for gender in genders:
            for _ in range(25):
                assignments.append((dept, gender))
    # Shuffle with the seeded random
    random.shuffle(assignments)

    for i in range(300):
        row = i + 2
        emp_id = f'EMP{1001 + i:04d}'
        dept, gender = assignments[i]
        sal_min, sal_max = salary_ranges[dept]
        salary = round(random.uniform(sal_min, sal_max), 2)
        years = random.randint(1, 25)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=dept)
        ws.cell(row=row, column=3, value=gender)
        ws.cell(row=row, column=4, value=salary)
        ws.cell(row=row, column=4).number_format = '#,##0.00'
        ws.cell(row=row, column=5, value=years)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
