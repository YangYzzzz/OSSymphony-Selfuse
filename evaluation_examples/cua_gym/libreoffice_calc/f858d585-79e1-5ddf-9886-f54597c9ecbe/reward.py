"""
Reward Script: Calculate employee tenure with DATEDIF and highlight 10+ years in gold
Task ID: calc_hr_employee_tenure_002
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.2): E1 header is 'Years of Service'
  - Component 2 (0.4): E2:E92 contain DATEDIF(Dx,TODAY(),"Y") formulas
  - Component 3 (0.4): Conditional formatting on A2:E92 with E2>=10 and gold fill (#FFD700)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_employee_tenure_002'


def check_cf_on_range(ws, target_range='A2:E92', gold_argb='FFFFD700'):
    """
    Check whether there is a conditional formatting rule on target_range
    that uses formula E2>=10 and applies a gold fill.
    Returns (range_ok, formula_ok, color_ok).
    """
    range_ok = False
    formula_ok = False
    color_ok = False

    for cf_range in ws.conditional_formatting:
        cf_range_str = str(cf_range)
        if target_range not in cf_range_str and cf_range_str != target_range:
            continue
        # Range found
        range_ok = True
        cf_list = ws.conditional_formatting[cf_range]
        for rule in cf_list:
            # Check formula
            if hasattr(rule, 'formula') and rule.formula:
                formula_str = ''.join(rule.formula).upper().replace(' ', '')
                if re.search(r'\$?E2>=10', formula_str):
                    formula_ok = True
            # Check gold fill
            if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                for attr in ('fgColor', 'bgColor'):
                    try:
                        rgb_val = getattr(rule.dxf.fill, attr).rgb
                        if rgb_val and rgb_val.upper() == gold_argb.upper():
                            color_ok = True
                    except Exception:
                        pass

    return range_ok, formula_ok, color_ok


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Access the 'Staff' sheet
    try:
        ws = wb['Staff']
    except KeyError:
        print("CRITICAL: Sheet 'Staff' not found")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: E1 header is 'Years of Service' (0.2 points)
    # This FAILS on initial (E1=None) → PASSES on golden (E1='Years of Service')
    try:
        e1_value = ws['E1'].value
        if e1_value is not None and str(e1_value).strip() == 'Years of Service':
            print(f"PASS: Component 1 — E1 header is 'Years of Service' (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 — expected E1='Years of Service', found: {repr(e1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E2:E92 contain DATEDIF formulas (0.4 points)
    # Each formula must match =DATEDIF(Dx,TODAY(),"Y") for the appropriate row number.
    # This FAILS on initial (all E2:E92 are None) → PASSES on golden (all have formulas)
    try:
        formula_count = 0
        formula_errors = []
        expected_rows = 91  # rows 2 through 92

        for row in range(2, 93):
            cell_value = ws.cell(row=row, column=5).value
            if cell_value is None:
                formula_errors.append(f"E{row} is empty")
                continue
            val_str = str(cell_value).strip().upper().replace(' ', '')
            # Pattern: =DATEDIF(D<row>,TODAY(),"Y")
            expected_ref = f'D{row}'
            pattern = re.compile(
                r'^=DATEDIF\(' + re.escape(expected_ref.upper()) + r',TODAY\(\),"Y"\)$'
            )
            if pattern.match(val_str):
                formula_count += 1
            else:
                formula_errors.append(f"E{row}={repr(cell_value)}")

        if formula_count == expected_rows:
            print(f"PASS: Component 2 — All {expected_rows} DATEDIF formulas present in E2:E92 (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — only {formula_count}/{expected_rows} DATEDIF formulas correct")
            if formula_errors:
                print(f"  First errors: {formula_errors[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on A2:E92 with formula $E2>=10 and gold fill (0.4 points)
    # This FAILS on initial (no conditional formatting) → PASSES on golden
    # Gold color: #FFD700 → ARGB = FFFFD700
    try:
        range_ok, formula_ok, color_ok = check_cf_on_range(ws, 'A2:E92', 'FFFFD700')

        if not range_ok:
            print("FAIL: Component 3 — No conditional formatting rules found on range A2:E92")
        elif not formula_ok:
            print("FAIL: Component 3 — CF formula does not reference E2>=10 condition")
        elif not color_ok:
            print("FAIL: Component 3 — CF fill color is not gold (#FFD700 / FFFFD700)")
        else:
            print("PASS: Component 3 — CF on A2:E92 with E2>=10 and gold fill FFFFD700 (0.4 pts)")
            total_score += 0.4

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
