"""
Initial Setup: HR Employee Tenure Spreadsheet
Task ID: calc_hr_employee_tenure_002
Domain: libreoffice_calc

Creates a spreadsheet with employee data on sheet 'Staff'.
Column E is intentionally empty (no header, no data) so the agent can add 'Years of Service'.
"""

import os
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_employee_tenure_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Staff'

    # --- Headers ---
    # A=Emp ID, B=Name, C=Department, D=Start Date, E=<empty>, F=Salary
    ws['A1'] = 'Emp ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Department'
    ws['D1'] = 'Start Date'
    # E1 is intentionally left blank (Years of Service goes here)
    ws['F1'] = 'Salary'

    # Style headers with bold
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        cell = ws[f'{col_letter}1']
        cell.font = Font(bold=True)

    # --- Employee data (91 rows, rows 2-92) ---
    departments = ['Engineering', 'Marketing', 'Finance', 'HR', 'Operations',
                   'Sales', 'Legal', 'Product', 'Design', 'Customer Support']

    first_names = [
        'Sarah', 'Marcus', 'Priya', 'James', 'Olivia', 'Diego', 'Mei', 'Tobias',
        'Aisha', 'Ryan', 'Elena', 'Samuel', 'Fatima', 'Lucas', 'Ingrid', 'Kwame',
        'Nadia', 'Patrick', 'Yuki', 'Charlotte', 'Omar', 'Rachel', 'Andre', 'Simone',
        'Derek', 'Layla', 'Viktor', 'Grace', 'Mateo', 'Hannah', 'Tariq', 'Chloe',
        'Finn', 'Adaeze', 'Jason', 'Leila', 'Boris', 'Tanya', 'Nguyen', 'Bridget',
        'Alejandro', 'Suki', 'Malcolm', 'Vera', 'Caleb', 'Amara', 'Stefan', 'Jess',
        'Rohan', 'Claire', 'Antoine', 'Sasha', 'Hiro', 'Miriam', 'Travis', 'Lydia',
        'Kofi', 'Anya', 'Brennan', 'Naomi', 'Ivan', 'Deja', 'Morgan', 'Felipe',
        'Astrid', 'Darius', 'Rina', 'Seamus', 'Yara', 'Conrad', 'Imani', 'Lars',
        'Valentina', 'Ethan', 'Zoe', 'Joaquin', 'Petra', 'Leon', 'Sanaa', 'Mikael',
        'Nkechi', 'Oscar', 'Jade', 'Henry', 'Fatou', 'Soren', 'Celia', 'Kasim',
        'Lily', 'Dmitri'
    ]

    last_names = [
        'Chen', 'Johnson', 'Patel', 'Williams', 'Torres', 'Nguyen', 'Fischer',
        'Okonkwo', 'Campbell', 'Morrison', 'Ramirez', 'Andersen', 'Hassan', 'Ferreira',
        'Lindqvist', 'Mensah', 'Popescu', 'Sullivan', 'Tanaka', 'Beaumont', 'Al-Rashid',
        'Kwan', 'Santos', 'Dubois', 'Henderson', 'Khalil', 'Petrov', 'Nakamura',
        'Vargas', 'Schmidt', 'Ibrahim', 'Larsson', 'Murphy', 'Eze', 'Park',
        'Farouk', 'Volkov', 'Orlov', 'Tran', 'Gallagher', 'Reyes', 'Hayashi',
        'Winters', 'Kuznetsov', 'Brooks', 'Diallo', 'Weber', 'Novak', 'Sharma',
        'Laurent', 'Morel', 'Ivanova', 'Yamamoto', 'Cohen', 'Becker', 'Sousa',
        'Asante', 'Babic', 'Quinn', 'Osei', 'Nikolaev', 'Morgan', 'Kim', 'Castro',
        'Berg', 'Adeyemi', 'Sato', 'Walsh', 'Abbas', 'Hoffmann', 'Nwosu', 'Eriksson',
        'Mendoza', 'Clarke', 'Herrera', 'Gomez', 'Novotny', 'De Bruyne', 'Diop',
        'Lindberg', 'Achebe', 'Svensson', 'Hudson', 'Kamara', 'Persson', 'Delgado',
        'Nkrumah', 'Johansson', 'White', 'Lebedev', 'Abubakar'
    ]

    # Start dates range from 2006-01-01 to 2023-12-31
    # Mix of old employees (10+ years) and recent ones
    start_dates = [
        date(2006, 3, 15), date(2007, 8, 1),  date(2008, 11, 20), date(2009, 4, 5),
        date(2010, 2, 14), date(2006, 6, 30),  date(2011, 9, 12), date(2007, 1, 22),
        date(2012, 5, 8),  date(2013, 7, 17),  date(2006, 10, 3), date(2014, 3, 25),
        date(2015, 1, 9),  date(2008, 12, 11), date(2016, 4, 18), date(2009, 2, 27),
        date(2017, 6, 6),  date(2010, 8, 14),  date(2018, 3, 21), date(2019, 11, 4),
        date(2006, 5, 19), date(2007, 9, 28),  date(2020, 1, 13), date(2011, 7, 7),
        date(2021, 2, 22), date(2008, 4, 16),  date(2022, 6, 10), date(2009, 11, 30),
        date(2023, 1, 5),  date(2010, 3, 8),   date(2006, 7, 24), date(2013, 5, 13),
        date(2014, 8, 29), date(2007, 6, 3),   date(2015, 10, 17), date(2012, 2, 9),
        date(2016, 12, 1), date(2008, 9, 18),  date(2017, 4, 26), date(2009, 7, 14),
        date(2006, 1, 11), date(2018, 8, 7),   date(2010, 12, 22), date(2019, 3, 16),
        date(2011, 6, 4),  date(2020, 9, 29),  date(2007, 3, 13), date(2021, 5, 20),
        date(2008, 7, 9),  date(2022, 10, 3),  date(2009, 1, 27), date(2023, 4, 14),
        date(2010, 5, 5),  date(2006, 9, 8),   date(2013, 11, 19), date(2014, 2, 1),
        date(2007, 10, 17), date(2015, 7, 25),  date(2012, 4, 12), date(2016, 9, 6),
        date(2008, 6, 21), date(2017, 11, 14),  date(2009, 3, 30), date(2018, 5, 23),
        date(2010, 10, 11), date(2019, 7, 8),  date(2006, 12, 16), date(2020, 4, 27),
        date(2011, 2, 3),  date(2021, 9, 15),  date(2007, 5, 22), date(2022, 3, 31),
        date(2008, 8, 10), date(2023, 6, 19),  date(2009, 10, 4), date(2006, 4, 29),
        date(2013, 2, 17), date(2014, 6, 8),   date(2010, 7, 26), date(2015, 3, 12),
        date(2012, 9, 23), date(2016, 11, 7),  date(2007, 12, 5), date(2017, 2, 18),
        date(2008, 3, 27), date(2018, 10, 2),  date(2009, 6, 15), date(2006, 8, 20),
        date(2019, 12, 11), date(2011, 4, 6),  date(2020, 7, 22)
    ]

    salaries = [
        72500, 85000, 91200, 64800, 110000, 78300, 95600, 68000, 120000, 58500,
        142000, 73200, 88700, 55000, 103500, 67900, 115000, 82400, 61000, 76800,
        98200, 145000, 53500, 87600, 71300, 112000, 59800, 94500, 80000, 138000,
        65700, 107000, 78900, 56200, 92800, 69400, 125000, 84100, 62300, 100500,
        74600, 118000, 57800, 89300, 81700, 133000, 66500, 104000, 79200, 54000,
        96700, 71800, 109000, 83500, 60400, 127000, 75100, 88000, 63200, 102000,
        77400, 116000, 58100, 91500, 82900, 141000, 67300, 105500, 80700, 55900,
        97900, 72800, 111000, 84800, 61600, 128000, 76300, 89500, 64100, 103000,
        78600, 117500, 59300, 92700, 83200, 139000, 68100, 106000, 81300, 56700,
        98400
    ]

    for i in range(91):
        row = i + 2
        emp_id = f'EMP{str(i + 1001).zfill(4)}'
        full_name = f'{first_names[i % len(first_names)]} {last_names[i % len(last_names)]}'
        dept = departments[i % len(departments)]
        start_date = start_dates[i]
        salary = salaries[i]

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=full_name)
        ws.cell(row=row, column=3, value=dept)
        ws.cell(row=row, column=4, value=start_date)
        ws.cell(row=row, column=4).number_format = 'yyyy-mm-dd'
        # Column E (5) intentionally left empty — for Years of Service
        ws.cell(row=row, column=6, value=salary)
        ws.cell(row=row, column=6).number_format = '$#,##0'

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Staff, Rows: 1 header + 91 data = 92 rows')
    print(f'Column E is empty (no header, no data) — ready for Years of Service')

create_initial()
