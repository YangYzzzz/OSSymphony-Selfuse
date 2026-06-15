"""
Reward Script: Use ISNUMBER and ISTEXT formulas to flag numeric and text entries
Task ID: calc_fma_isnumber_istext_042
Domain: libreoffice_calc
Scoring:
  - Component 1: ISNUMBER formulas in B2:B13 (0.5 pts)
  - Component 2: ISTEXT formulas in C2:C13 (0.5 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_isnumber_istext_042'
SHEET_NAME = 'DataAudit'
DATA_ROWS = range(2, 14)  # rows 2 through 13


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Column A contains mixed numbers and text values.
    Agent must add ISNUMBER formulas in B2:B13 and ISTEXT formulas in C2:C13.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the target sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: ISNUMBER formulas in B2:B13 (0.5 points)
    # Each row should have a formula =ISNUMBER(Ax) in column B
    try:
        isnumber_correct = 0
        isnumber_total = len(DATA_ROWS)
        isnumber_failures = []

        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=2).value  # Column B
            expected_formula = f"=ISNUMBER(A{row})"

            if cell_val is None:
                isnumber_failures.append(f"B{row}: empty (expected {expected_formula})")
                continue

            if isinstance(cell_val, str):
                # Normalize: strip whitespace, uppercase for comparison
                normalized = cell_val.strip().upper().replace(" ", "")
                expected_norm = expected_formula.upper().replace(" ", "")
                if normalized == expected_norm:
                    isnumber_correct += 1
                else:
                    isnumber_failures.append(f"B{row}: got {repr(cell_val)}, expected {expected_formula}")
            else:
                isnumber_failures.append(f"B{row}: got non-formula value {repr(cell_val)}")

        if isnumber_correct == isnumber_total:
            print(f"PASS: Component 1 — All {isnumber_total} ISNUMBER formulas present in B2:B13 (0.5 pts)")
            total_score += 0.5
        elif isnumber_correct > 0:
            # Partial credit: at least some formulas present
            partial = round(0.5 * isnumber_correct / isnumber_total, 4)
            print(f"PARTIAL: Component 1 — {isnumber_correct}/{isnumber_total} ISNUMBER formulas correct ({partial} pts)")
            print(f"  Failures: {isnumber_failures[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No ISNUMBER formulas found in B2:B13 (0.0 pts)")
            if isnumber_failures:
                print(f"  Sample failures: {isnumber_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 (ISNUMBER check) — {e}")

    # Component 2: ISTEXT formulas in C2:C13 (0.5 points)
    # Each row should have a formula =ISTEXT(Ax) in column C
    try:
        istext_correct = 0
        istext_total = len(DATA_ROWS)
        istext_failures = []

        for row in DATA_ROWS:
            cell_val = ws.cell(row=row, column=3).value  # Column C
            expected_formula = f"=ISTEXT(A{row})"

            if cell_val is None:
                istext_failures.append(f"C{row}: empty (expected {expected_formula})")
                continue

            if isinstance(cell_val, str):
                # Normalize: strip whitespace, uppercase for comparison
                normalized = cell_val.strip().upper().replace(" ", "")
                expected_norm = expected_formula.upper().replace(" ", "")
                if normalized == expected_norm:
                    istext_correct += 1
                else:
                    istext_failures.append(f"C{row}: got {repr(cell_val)}, expected {expected_formula}")
            else:
                istext_failures.append(f"C{row}: got non-formula value {repr(cell_val)}")

        if istext_correct == istext_total:
            print(f"PASS: Component 2 — All {istext_total} ISTEXT formulas present in C2:C13 (0.5 pts)")
            total_score += 0.5
        elif istext_correct > 0:
            # Partial credit: at least some formulas present
            partial = round(0.5 * istext_correct / istext_total, 4)
            print(f"PARTIAL: Component 2 — {istext_correct}/{istext_total} ISTEXT formulas correct ({partial} pts)")
            print(f"  Failures: {istext_failures[:5]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No ISTEXT formulas found in C2:C13 (0.0 pts)")
            if istext_failures:
                print(f"  Sample failures: {istext_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 (ISTEXT check) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
