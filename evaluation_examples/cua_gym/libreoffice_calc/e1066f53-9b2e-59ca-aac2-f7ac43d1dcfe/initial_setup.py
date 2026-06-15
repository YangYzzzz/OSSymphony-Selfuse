"""
Initial Setup: DataAudit spreadsheet with mixed values in Column A
Task ID: calc_fma_isnumber_istext_042
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_isnumber_istext_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DataAudit ---
    ws = wb.active
    ws.title = 'DataAudit'

    # Headers in row 1
    ws['A1'] = 'Value'
    ws['B1'] = 'Is Number?'
    ws['C1'] = 'Is Text?'

    # Mixed values in A2:A13 as described in context
    # Note: '2024' is a string (imported from external system as text)
    # TRUE is a boolean, 0 is a number
    mixed_values = [
        123,       # numeric
        'ABC',     # text
        45.6,      # numeric
        'hello',   # text
        789,       # numeric
        '2024',    # text (string, not number)
        True,      # boolean
        'xyz',     # text
        0,         # numeric
        'test',    # text
        99.9,      # numeric
        'N/A',     # text
    ]

    for row_idx, value in enumerate(mixed_values, start=2):
        ws.cell(row=row_idx, column=1, value=value)
        # Columns B and C are intentionally left empty (no formulas)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: DataAudit')
    print('Column A (rows 2-13): Mixed values (numbers, strings, boolean)')
    print('Column B (rows 2-13): Empty (Is Number? to be filled by agent)')
    print('Column C (rows 2-13): Empty (Is Text? to be filled by agent)')


create_initial()
