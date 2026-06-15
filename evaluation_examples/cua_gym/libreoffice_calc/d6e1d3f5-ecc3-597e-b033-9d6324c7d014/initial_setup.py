"""
Initial Setup: Employee performance review scores - raw data only.
Task ID: calc_gpm_023
Domain: libreoffice_calc

Creates PerfReview sheet with employee names and scores (B-F),
but NO formulas in G/H, NO conditional formatting, NO chart.
The agent's task is to add those elements.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_023'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PerfReview'

    # --- Row 1: Title (merged A1:H1) ---
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'Annual Performance Review Summary - 2025'
    title_cell.font = Font(size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='FF006666', end_color='FF006666', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Row 3: Headers ---
    headers = ['Employee', 'Technical', 'Communication', 'Leadership',
               'Initiative', 'Teamwork', 'Overall', 'Rating']
    header_fill = PatternFill(start_color='FF006666', end_color='FF006666', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center')

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Rows 4-9: Employee data (scores in B-F only, G and H left empty) ---
    employees = [
        ['Alice',  4.5, 3.8, 4.0, 4.2, 4.5],
        ['Bob',    3.5, 4.2, 3.0, 3.8, 4.0],
        ['Carol',  4.8, 4.5, 4.2, 4.0, 3.5],
        ['Dan',    3.0, 3.5, 4.5, 3.2, 4.2],
        ['Eve',    4.2, 4.0, 3.8, 4.5, 4.0],
        ['Frank',  3.8, 3.2, 3.5, 3.0, 3.8],
    ]

    for r_idx, emp in enumerate(employees, 4):
        # Column A: Employee name
        ws.cell(row=r_idx, column=1, value=emp[0])
        # Columns B-F: Scores
        for c_idx, score in enumerate(emp[1:], 2):
            cell = ws.cell(row=r_idx, column=c_idx, value=score)
            cell.number_format = '0.0'
            cell.alignment = Alignment(horizontal='center')

    # --- Borders on all data cells (rows 3-9, cols A-H) ---
    thin = Side(style='thin', color='000000')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=3, max_row=9, min_col=1, max_col=8):
        for cell in row:
            cell.border = border_all

    # --- Column widths for readability ---
    ws.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
