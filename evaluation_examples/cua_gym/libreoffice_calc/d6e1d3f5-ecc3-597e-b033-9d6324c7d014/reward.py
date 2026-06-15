"""
Reward Script: Employee Performance Review with radar-style comparison table
Task ID: calc_gpm_023
Domain: libreoffice_calc
Scoring:
  Component 1: AVERAGE formulas in G4:G9 (0.25 pts)
  Component 2: IF rating formulas in H4:H9 (0.25 pts)
  Component 3: Color scale conditional formatting on B4:F9 (0.15 pts)
  Component 4: Rating conditional formatting on H4:H9 (0.15 pts)
  Component 5: Competency comparison chart present (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_023'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: PerfReview sheet must exist
    if 'PerfReview' not in wb.sheetnames:
        print("FAIL: Sheet 'PerfReview' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PerfReview']

    # Component 1: AVERAGE formulas in G4:G9 (0.25 points)
    # Initial env has G4:G9 as None; golden has =AVERAGE(B:F) formulas
    try:
        avg_count = 0
        for row in range(4, 10):
            val = ws.cell(row=row, column=7).value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                if "AVERAGE" in normalized and "B" in normalized and "F" in normalized:
                    avg_count += 1
        if avg_count == 6:
            print(f"PASS: Component 1 -- All 6 AVERAGE formulas found in G4:G9 (0.25 pts)")
            total_score += 0.25
        elif avg_count >= 3:
            partial = round(0.25 * avg_count / 6, 2)
            print(f"PARTIAL: Component 1 -- {avg_count}/6 AVERAGE formulas found in G4:G9 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 -- Only {avg_count}/6 AVERAGE formulas found in G4:G9")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: IF rating formulas in H4:H9 (0.25 points)
    # Initial env has H4:H9 as None; golden has nested IF formulas
    try:
        if_count = 0
        for row in range(4, 10):
            val = ws.cell(row=row, column=8).value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                if "IF" in normalized and "EXCEEDS" in normalized.upper() or "MEETS" in normalized.upper():
                    if_count += 1
        if if_count == 6:
            print(f"PASS: Component 2 -- All 6 IF rating formulas found in H4:H9 (0.25 pts)")
            total_score += 0.25
        elif if_count >= 3:
            partial = round(0.25 * if_count / 6, 2)
            print(f"PARTIAL: Component 2 -- {if_count}/6 IF formulas found in H4:H9 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 -- Only {if_count}/6 IF rating formulas found in H4:H9")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Color scale conditional formatting on B4:F9 (0.15 points)
    # Initial env has no conditional formatting; golden has 3-color scale on B4:F9
    try:
        has_color_scale = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            for rule in cf.rules:
                if rule.type == 'colorScale':
                    # Check that the range covers at least B4:F9
                    if 'B4' in cf_range or 'B4:F9' in cf_range:
                        has_color_scale = True
                        break
            if has_color_scale:
                break
        if has_color_scale:
            print(f"PASS: Component 3 -- Color scale conditional formatting found on score range (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 -- No color scale conditional formatting found on B4:F9")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Rating conditional formatting on H4:H9 (0.15 points)
    # Initial env has no conditional formatting; golden has cellIs rules on H4:H9
    try:
        rating_cf_count = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'H4' in cf_range or 'H4:H9' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        rating_cf_count += 1
        if rating_cf_count >= 2:
            print(f"PASS: Component 4 -- Rating conditional formatting found ({rating_cf_count} cellIs rules) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 -- Expected >=2 cellIs conditional formatting rules on H4:H9, found {rating_cf_count}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Chart present with title 'Competency Comparison' (0.20 points)
    # Initial env has 0 charts; golden has 1 chart
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Check for a chart with title containing "Competency Comparison"
            found_chart = False
            for ch in charts:
                chart_title = ''
                if ch.title is not None:
                    # Title can be a string or a Title object
                    if isinstance(ch.title, str):
                        chart_title = ch.title
                    else:
                        # Extract text from Title object
                        try:
                            if hasattr(ch.title, 'text'):
                                chart_title = ch.title.text or ''
                            if not chart_title and hasattr(ch.title, 'tx') and ch.title.tx:
                                if hasattr(ch.title.tx, 'rich') and ch.title.tx.rich:
                                    for p in ch.title.tx.rich.p:
                                        for r in p.r:
                                            chart_title += r.t or ''
                        except Exception:
                            chart_title = str(ch.title)

                chart_title = str(chart_title)
                if 'competency' in chart_title.lower() and 'comparison' in chart_title.lower():
                    found_chart = True
                    series_count = len(ch.series)
                    print(f"PASS: Component 5 -- Chart 'Competency Comparison' found with {series_count} series (0.20 pts)")
                    total_score += 0.20
                    break

            if not found_chart:
                # Still give partial credit if chart exists but title differs
                series_count = len(charts[0].series)
                if series_count >= 3:
                    print(f"PARTIAL: Component 5 -- Chart found with {series_count} series but title mismatch (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 5 -- Chart found but wrong title and only {series_count} series")
        else:
            print(f"FAIL: Component 5 -- No charts found in worksheet")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
