"""
Initial Setup: Headcount Cost Analysis - Total Cost of Employment
Task ID: calc_fin_headcount_cost_035
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_headcount_cost_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HeadcountCost'

    # Row 1 headers: Employee (A), Level (B), Base Salary (C)
    ws['A1'] = 'Employee'
    ws['B1'] = 'Level'
    ws['C1'] = 'Base Salary'

    # Assumption cells as specified in context
    # G1=0.15 (benefits rate), G2=0.0765 (FICA), G3=0.05 (equity rate), G4=0.10 (overhead rate)
    ws['G1'] = 0.15
    ws['G2'] = 0.0765
    ws['G3'] = 0.05
    ws['G4'] = 0.10

    # Labels for assumption cells in column H
    ws['H1'] = 'Benefits Rate'
    ws['H2'] = 'FICA Rate'
    ws['H3'] = 'Equity Rate'
    ws['H4'] = 'Overhead Rate'

    # Realistic employee data: 29 rows (rows 2-30)
    # Levels: Junior, Mid, Senior, Lead, Director
    employees = [
        ('Sarah Chen', 'Senior', 95000),
        ('Marcus Johnson', 'Mid', 72000),
        ('Emily Rodriguez', 'Lead', 115000),
        ('David Park', 'Junior', 58000),
        ('Jessica Williams', 'Senior', 92000),
        ('Robert Kim', 'Director', 145000),
        ('Amanda Foster', 'Mid', 75000),
        ('Christopher Lee', 'Senior', 98000),
        ('Nicole Thompson', 'Junior', 55000),
        ('James Martinez', 'Lead', 118000),
        ('Rachel Anderson', 'Mid', 78000),
        ('Brandon White', 'Senior', 102000),
        ('Stephanie Harris', 'Director', 152000),
        ('Kevin Brown', 'Junior', 60000),
        ('Lauren Garcia', 'Mid', 71000),
        ('Michael Davis', 'Lead', 122000),
        ('Tiffany Wilson', 'Senior', 96000),
        ('Andrew Taylor', 'Junior', 57000),
        ('Samantha Moore', 'Mid', 74000),
        ('Jason Jackson', 'Lead', 125000),
        ('Melissa Thomas', 'Senior', 99000),
        ('Ryan Hernandez', 'Director', 148000),
        ('Angela Martin', 'Mid', 77000),
        ('Daniel Thompson', 'Junior', 62000),
        ('Jennifer Garcia', 'Senior', 94000),
        ('Patrick White', 'Lead', 119000),
        ('Christina Lewis', 'Mid', 73000),
        ('Steven Robinson', 'Junior', 56000),
        ('Vanessa Clark', 'Senior', 101000),
    ]

    for r, (emp, level, salary) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp)
        ws.cell(row=r, column=2, value=level)
        ws.cell(row=r, column=3, value=salary)

    # Column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: HeadcountCost')
    print(f'  Rows: 1 header + 29 data rows')
    print(f'  Columns: A (Employee), B (Level), C (Base Salary)')
    print(f'  Assumptions: G1=0.15 (benefits), G2=0.0765 (FICA), G3=0.05 (equity), G4=0.10 (overhead)')


create_initial()
