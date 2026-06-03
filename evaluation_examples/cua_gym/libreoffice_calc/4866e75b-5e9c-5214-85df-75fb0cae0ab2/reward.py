"""
Reward Script: Total Cost of Employment (TCE) Analysis
Task ID: calc_fin_headcount_cost_035
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: New column headers added (D1:K1)                    — 0.15 pts
  Component 2: Cost component formulas in D:G cols (rows 2-30)     — 0.25 pts
  Component 3: Total TCE (SUM) formulas in H col (rows 2-30)       — 0.15 pts
  Component 4: Percentage formulas in I:K cols (rows 2-30)         — 0.15 pts
  Component 5: Currency number format on C:H, % format on I:K      — 0.10 pts
  Component 6: Row 1 bold formatting + freeze panes at A2           — 0.10 pts
  Component 7: Stacked bar chart exists                             — 0.10 pts
  Total: 1.00
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_headcount_cost_035'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'HeadcountCost' must exist
    if 'HeadcountCost' not in wb.sheetnames:
        print("CRITICAL: Sheet 'HeadcountCost' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['HeadcountCost']

    # -------------------------------------------------------------------------
    # Component 1: New column headers added (D1:K1) — 0.15 points
    # The initial file has no headers in columns D-K. The golden file should have:
    # D1='Benefits', E1='FICA', F1='Equity', G1='Overhead', H1='Total TCE',
    # I1='Base %', J1='Benefits %', K1='Other %'
    # -------------------------------------------------------------------------
    try:
        expected_headers = {
            4: 'Benefits',
            5: 'FICA',
            6: 'Equity',
            7: 'Overhead',
            8: 'Total TCE',
            9: 'Base %',
            10: 'Benefits %',
            11: 'Other %',
        }
        headers_found = 0
        for col_idx, expected_name in expected_headers.items():
            actual_val = ws.cell(row=1, column=col_idx).value
            if actual_val and str(actual_val).strip().lower() == expected_name.lower():
                headers_found += 1
            else:
                print(f"FAIL: Header col {col_idx} — expected '{expected_name}', found: {repr(actual_val)}")

        if headers_found == len(expected_headers):
            print(f"PASS: Component 1 — All 8 new headers present (0.15 pts)")
            total_score += 0.15
        elif headers_found >= 5:
            print(f"PARTIAL: Component 1 — {headers_found}/8 headers present (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 1 — Only {headers_found}/8 headers present")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Cost component formulas in D:G columns (rows 2-30) — 0.25 pts
    # Each cost formula should multiply Base Salary (col C) by an assumption rate.
    # The formulas reference column M (shifted assumption column) with $M$1-$M$4.
    # Pattern: D row = C*rate1, E row = C*rate2, F row = C*rate3, G row = C*rate4
    # -------------------------------------------------------------------------
    try:
        formula_pattern = re.compile(r'^=C\d+\s*\*\s*\$[A-Z]+\$[1-4]$', re.IGNORECASE)
        cost_cols = [4, 5, 6, 7]  # D, E, F, G
        total_checks = len(cost_cols) * 29  # 4 cols * 29 rows (rows 2-30)
        passed_checks = 0
        sampled_failures = []

        for row in range(2, 31):
            for col_idx in cost_cols:
                val = ws.cell(row=row, column=col_idx).value
                if val and isinstance(val, str):
                    # Accept formulas like =C2*$M$1, =C2*$G$1, or similar
                    cleaned = val.replace(' ', '').upper()
                    # Must multiply column C by an absolute rate reference
                    if re.match(r'^=C\d+\*\$[A-Z]+\$\d+$', cleaned):
                        passed_checks += 1
                    else:
                        if len(sampled_failures) < 3:
                            sampled_failures.append(f"Row {row} Col {col_idx}: {repr(val)}")
                else:
                    if len(sampled_failures) < 3:
                        sampled_failures.append(f"Row {row} Col {col_idx}: {repr(val)} (not a formula)")

        ratio = passed_checks / total_checks
        if ratio >= 0.95:
            print(f"PASS: Component 2 — {passed_checks}/{total_checks} cost formulas present (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = round(0.25 * ratio, 2)
            print(f"PARTIAL: Component 2 — {passed_checks}/{total_checks} cost formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {passed_checks}/{total_checks} cost formulas. Failures: {sampled_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Total TCE (SUM) formulas in H column (rows 2-30) — 0.15 pts
    # H col should have =SUM(C{row}:G{row}) for each data row.
    # -------------------------------------------------------------------------
    try:
        sum_formula_pattern = re.compile(r'^=SUM\(C\d+:G\d+\)$', re.IGNORECASE)
        sum_passed = 0
        sum_failures = []

        for row in range(2, 31):
            val = ws.cell(row=row, column=8).value  # H column
            if val and isinstance(val, str):
                cleaned = val.replace(' ', '').upper()
                if re.match(r'^=SUM\(C\d+:G\d+\)$', cleaned):
                    sum_passed += 1
                else:
                    if len(sum_failures) < 3:
                        sum_failures.append(f"H{row}: {repr(val)}")
            else:
                if len(sum_failures) < 3:
                    sum_failures.append(f"H{row}: {repr(val)}")

        if sum_passed == 29:
            print(f"PASS: Component 3 — All 29 SUM formulas in H column (0.15 pts)")
            total_score += 0.15
        elif sum_passed >= 15:
            partial = round(0.15 * (sum_passed / 29), 2)
            print(f"PARTIAL: Component 3 — {sum_passed}/29 SUM formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {sum_passed}/29 SUM formulas. Failures: {sum_failures}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Percentage formulas in I:K columns (rows 2-30) — 0.15 pts
    # I col: =C{row}/H{row} (Base %)
    # J col: =D{row}/H{row} (Benefits %)
    # K col: formula involving (E+F+G)/H (Other %)
    # -------------------------------------------------------------------------
    try:
        pct_passed = 0
        pct_total = 29 * 3  # 3 cols * 29 rows

        for row in range(2, 31):
            # Check I column: Base % = C/H
            val_i = ws.cell(row=row, column=9).value
            if val_i and isinstance(val_i, str):
                cleaned_i = val_i.replace(' ', '').upper()
                if re.match(r'^=C\d+/H\d+$', cleaned_i):
                    pct_passed += 1

            # Check J column: Benefits % = D/H
            val_j = ws.cell(row=row, column=10).value
            if val_j and isinstance(val_j, str):
                cleaned_j = val_j.replace(' ', '').upper()
                if re.match(r'^=D\d+/H\d+$', cleaned_j):
                    pct_passed += 1

            # Check K column: Other % = some combination of E,F,G / H
            val_k = ws.cell(row=row, column=11).value
            if val_k and isinstance(val_k, str):
                cleaned_k = val_k.replace(' ', '').upper()
                # Should reference H in denominator and E/F/G in numerator
                if '/H' in cleaned_k and ('E' in cleaned_k or 'F' in cleaned_k):
                    pct_passed += 1

        ratio = pct_passed / pct_total
        if ratio >= 0.95:
            print(f"PASS: Component 4 — {pct_passed}/{pct_total} percentage formulas (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.5:
            partial = round(0.15 * ratio, 2)
            print(f"PARTIAL: Component 4 — {pct_passed}/{pct_total} percentage formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {pct_passed}/{pct_total} percentage formulas")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------------
    # Component 5: Currency format on C:H, percentage format on I:K — 0.10 pts
    # Currency columns (C-H): should contain '$' or currency pattern
    # Percentage columns (I-K): should contain '%' pattern
    # -------------------------------------------------------------------------
    try:
        currency_cols_ok = 0
        pct_cols_ok = 0
        currency_cols = [3, 4, 5, 6, 7, 8]  # C through H
        pct_cols = [9, 10, 11]  # I through K

        for col_idx in currency_cols:
            nf = ws.cell(row=2, column=col_idx).number_format
            if nf and ('$' in nf or '#,##0' in nf.upper() or 'currency' in nf.lower()):
                currency_cols_ok += 1
            else:
                print(f"FAIL: C5 currency check — col {col_idx} format: {repr(nf)}")

        for col_idx in pct_cols:
            nf = ws.cell(row=2, column=col_idx).number_format
            if nf and '%' in nf:
                pct_cols_ok += 1
            else:
                print(f"FAIL: C5 pct check — col {col_idx} format: {repr(nf)}")

        total_format_checks = len(currency_cols) + len(pct_cols)
        total_format_passed = currency_cols_ok + pct_cols_ok

        if total_format_passed == total_format_checks:
            print(f"PASS: Component 5 — All number formats correct (0.10 pts)")
            total_score += 0.10
        elif total_format_passed >= 5:
            partial = round(0.10 * (total_format_passed / total_format_checks), 2)
            print(f"PARTIAL: Component 5 — {total_format_passed}/{total_format_checks} formats correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {total_format_passed}/{total_format_checks} formats correct")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # -------------------------------------------------------------------------
    # Component 6: Row 1 bold formatting + freeze panes at A2 — 0.10 pts
    # Row 1 headers should be bold (was not bold in initial file).
    # Freeze panes should be at A2 (was None in initial file).
    # -------------------------------------------------------------------------
    try:
        # Check bold for all new header cells D1 to K1 (and at least A1)
        bold_cells = [1, 4, 5, 6, 7, 8, 9, 10, 11]  # A1, D1-K1
        bold_count = sum(1 for col in bold_cells if ws.cell(row=1, column=col).font.bold)
        bold_ok = bold_count >= len(bold_cells) * 0.7  # at least 70% bold

        # Check freeze panes
        freeze_ok = ws.freeze_panes is not None and str(ws.freeze_panes).upper() in ['A2', 'B2']

        if bold_ok and freeze_ok:
            print(f"PASS: Component 6 — Row 1 bold ({bold_count}/{len(bold_cells)} cells) and freeze panes at {ws.freeze_panes} (0.10 pts)")
            total_score += 0.10
        elif bold_ok or freeze_ok:
            print(f"PARTIAL: Component 6 — bold={bold_ok} (count={bold_count}), freeze={freeze_ok} ({ws.freeze_panes}) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — bold count={bold_count}/{len(bold_cells)}, freeze_panes={ws.freeze_panes}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # -------------------------------------------------------------------------
    # Component 7: Stacked bar chart exists — 0.10 pts
    # The initial file has 0 charts. Golden should have a stacked bar chart
    # showing cost components by employee level.
    # -------------------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Verify it's a stacked bar/column chart
            chart = charts[0]
            chart_type = type(chart).__name__
            grouping = getattr(chart, 'grouping', None)
            series_count = len(chart.series)

            if 'Bar' in chart_type or 'bar' in chart_type.lower():
                if grouping and 'stacked' in str(grouping).lower():
                    if series_count >= 2:
                        print(f"PASS: Component 7 — Stacked bar chart found with {series_count} series (0.10 pts)")
                        total_score += 0.10
                    else:
                        print(f"PARTIAL: Component 7 — Bar chart found but only {series_count} series (0.05 pts)")
                        total_score += 0.05
                else:
                    print(f"PARTIAL: Component 7 — Chart found (type={chart_type}) but grouping={grouping}, not stacked (0.05 pts)")
                    total_score += 0.05
            else:
                print(f"PARTIAL: Component 7 — Chart exists but type={chart_type}, expected BarChart (0.05 pts)")
                total_score += 0.05
        else:
            print(f"FAIL: Component 7 — No charts found (expected stacked bar chart)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
