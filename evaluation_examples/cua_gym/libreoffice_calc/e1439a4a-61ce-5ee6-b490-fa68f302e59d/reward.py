"""
Reward Script: Apply 'Good'/'Bad' cell styles to performance tracker
Task ID: calc_gfl_040
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): 'Good' style applied to 'Above Target' cells (green fill + font)
  - Component 2 (0.4): 'Bad' style applied to 'Below Target' cells (red fill + font)
  - Component 3 (0.2): All 24 cells in E2:E25 have SOME style applied (no unstyled cells)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_040'

# 'Good' style colors (LibreOffice built-in)
GOOD_FILL = 'FFC6EFCE'   # light green background
GOOD_FONT = '00006100'   # dark green font (may vary slightly; also check prefix patterns)

# 'Bad' style colors (LibreOffice built-in)
BAD_FILL = 'FFFFC7CE'    # light red/pink background
BAD_FONT = '009C0006'    # dark red font


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Targets' sheet must exist
    if 'Targets' not in wb.sheetnames:
        print("FAIL: 'Targets' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Targets']

    # Collect cells by status
    above_cells = []
    below_cells = []
    for r in range(2, 26):
        val = ws.cell(row=r, column=5).value
        if val is not None:
            val_str = str(val).strip()
            if 'above' in val_str.lower():
                above_cells.append(r)
            elif 'below' in val_str.lower():
                below_cells.append(r)

    if not above_cells and not below_cells:
        print("FAIL: No 'Above Target' or 'Below Target' values found in E2:E25")
        print("REWARD: 0.0")
        return 0.0

    # Helper to check if a cell has a "Good"-style fill (green-ish)
    def has_good_style(row):
        cell = ws.cell(row=row, column=5)
        try:
            fill_type = cell.fill.fill_type
            if fill_type != 'solid':
                return False
            fg = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            if fg is None:
                return False
            # Check for Good style green fill: FFC6EFCE
            # Also accept slight variants: the key is it should be a green-ish fill
            if fg == GOOD_FILL:
                return True
            # Accept if it's clearly a green fill (C6EFCE pattern)
            if 'C6EFCE' in fg:
                return True
            return False
        except Exception:
            return False

    # Helper to check if a cell has a "Bad"-style fill (red-ish)
    def has_bad_style(row):
        cell = ws.cell(row=row, column=5)
        try:
            fill_type = cell.fill.fill_type
            if fill_type != 'solid':
                return False
            fg = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            if fg is None:
                return False
            # Check for Bad style red fill: FFFFC7CE
            if fg == BAD_FILL:
                return True
            if 'FFC7CE' in fg:
                return True
            return False
        except Exception:
            return False

    # Helper to check if a cell has ANY non-default fill
    def has_any_fill(row):
        cell = ws.cell(row=row, column=5)
        try:
            fill_type = cell.fill.fill_type
            if fill_type is None or fill_type == 'none':
                return False
            fg = cell.fill.fgColor.rgb if cell.fill.fgColor else None
            if fg is None or fg == '00000000':
                return False
            return True
        except Exception:
            return False

    # Component 1: 'Good' style on 'Above Target' cells (0.4 points)
    try:
        good_pass = 0
        good_total = len(above_cells)
        for r in above_cells:
            if has_good_style(r):
                good_pass += 1
            else:
                cell = ws.cell(row=r, column=5)
                fg = None
                try:
                    fg = cell.fill.fgColor.rgb
                except:
                    pass
                print(f"  DETAIL: E{r} ('Above Target') missing Good style, fg={fg}")

        if good_total > 0:
            ratio = good_pass / good_total
            pts = round(0.4 * ratio, 4)
            if ratio == 1.0:
                print(f"PASS: Component 1 — All {good_total} 'Above Target' cells have 'Good' style (0.4 pts)")
                total_score += pts
            elif good_pass > 0:
                print(f"PARTIAL: Component 1 — {good_pass}/{good_total} 'Above Target' cells have 'Good' style ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 1 — 0/{good_total} 'Above Target' cells have 'Good' style")
        else:
            print("SKIP: Component 1 — No 'Above Target' cells found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Bad' style on 'Below Target' cells (0.4 points)
    try:
        bad_pass = 0
        bad_total = len(below_cells)
        for r in below_cells:
            if has_bad_style(r):
                bad_pass += 1
            else:
                cell = ws.cell(row=r, column=5)
                fg = None
                try:
                    fg = cell.fill.fgColor.rgb
                except:
                    pass
                print(f"  DETAIL: E{r} ('Below Target') missing Bad style, fg={fg}")

        if bad_total > 0:
            ratio = bad_pass / bad_total
            pts = round(0.4 * ratio, 4)
            if ratio == 1.0:
                print(f"PASS: Component 2 — All {bad_total} 'Below Target' cells have 'Bad' style (0.4 pts)")
                total_score += pts
            elif bad_pass > 0:
                print(f"PARTIAL: Component 2 — {bad_pass}/{bad_total} 'Below Target' cells have 'Bad' style ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 2 — 0/{bad_total} 'Below Target' cells have 'Bad' style")
        else:
            print("SKIP: Component 2 — No 'Below Target' cells found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All E2:E25 cells have some styling applied (0.2 points)
    try:
        styled_count = 0
        total_cells = 0
        for r in range(2, 26):
            val = ws.cell(row=r, column=5).value
            if val is not None:
                total_cells += 1
                if has_any_fill(r):
                    styled_count += 1

        if total_cells > 0:
            ratio = styled_count / total_cells
            pts = round(0.2 * ratio, 4)
            if ratio == 1.0:
                print(f"PASS: Component 3 — All {total_cells} cells in E2:E25 have fill styling (0.2 pts)")
                total_score += pts
            elif styled_count > 0:
                print(f"PARTIAL: Component 3 — {styled_count}/{total_cells} cells styled ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 3 — 0/{total_cells} cells in E2:E25 have fill styling")
        else:
            print("SKIP: Component 3 — No data cells found in E2:E25")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
