"""
Initial Setup: Apply Good/Bad cell styles to performance tracker
Task ID: calc_gfl_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Targets ---
    ws = wb.active
    ws.title = 'Targets'

    # Headers
    headers = ['Division', 'Month', 'Target', 'Actual', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 24 monthly performance entries (rows 2-25)
    data = [
        ['Northeast Sales', 'Jan 2025', 125000, 131500, 'Above Target'],
        ['Northeast Sales', 'Feb 2025', 128000, 119200, 'Below Target'],
        ['Southwest Ops', 'Jan 2025', 98000, 102400, 'Above Target'],
        ['Southwest Ops', 'Feb 2025', 101000, 94800, 'Below Target'],
        ['Central Logistics', 'Jan 2025', 87500, 91200, 'Above Target'],
        ['Central Logistics', 'Feb 2025', 89000, 85600, 'Below Target'],
        ['Pacific Marketing', 'Jan 2025', 145000, 152300, 'Above Target'],
        ['Pacific Marketing', 'Feb 2025', 148000, 141700, 'Below Target'],
        ['Atlantic Finance', 'Jan 2025', 76000, 78900, 'Above Target'],
        ['Atlantic Finance', 'Feb 2025', 78500, 72100, 'Below Target'],
        ['Mountain R&D', 'Jan 2025', 112000, 118700, 'Above Target'],
        ['Mountain R&D', 'Feb 2025', 115000, 109300, 'Below Target'],
        ['Great Lakes HR', 'Jan 2025', 65000, 68400, 'Above Target'],
        ['Great Lakes HR', 'Feb 2025', 67000, 63200, 'Below Target'],
        ['Southeast Supply', 'Jan 2025', 93000, 97600, 'Above Target'],
        ['Southeast Supply', 'Feb 2025', 95500, 88900, 'Below Target'],
        ['Northwest IT', 'Jan 2025', 134000, 140200, 'Above Target'],
        ['Northwest IT', 'Feb 2025', 137000, 129500, 'Below Target'],
        ['Midwest Production', 'Jan 2025', 156000, 163800, 'Above Target'],
        ['Midwest Production', 'Feb 2025', 159000, 151200, 'Below Target'],
        ['Northeast Sales', 'Mar 2025', 130000, 137400, 'Above Target'],
        ['Southwest Ops', 'Mar 2025', 103000, 96500, 'Below Target'],
        ['Central Logistics', 'Mar 2025', 91000, 95800, 'Above Target'],
        ['Pacific Marketing', 'Mar 2025', 151000, 144200, 'Below Target'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
