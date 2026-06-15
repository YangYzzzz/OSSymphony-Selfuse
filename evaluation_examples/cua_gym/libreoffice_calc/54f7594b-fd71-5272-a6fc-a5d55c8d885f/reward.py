"""
Reward Script: Assign inventory reorder priority levels using VLOOKUP
Task ID: osworld_calc_vlookup_grade_lookup_010
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): Column C has VLOOKUP approximate-match formulas for priority (rows 2-16)
  Component 2 (0.3 pts): Column G has VLOOKUP exact-match formulas for reorder quantity (rows 2-16)
  Component 3 (0.2 pts): Both columns have complete coverage across all 15 product rows
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_grade_lookup_010'

# Product data rows (rows 2 through 16, 15 products total)
PRODUCT_ROWS = range(2, 17)
TOTAL_ROWS = 15


def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, no spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def is_vlookup_approximate(formula_str):
    """
    Return True if formula_str is a VLOOKUP formula ending with ,1) or ,TRUE).
    Approximate match (last arg=1 or TRUE) is required for range-based lookups.
    """
    if not formula_str or not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    if not f.startswith('=VLOOKUP('):
        return False
    return f.endswith(',1)') or f.endswith(',TRUE)')


def is_vlookup_exact(formula_str):
    """
    Return True if formula_str is a VLOOKUP formula ending with ,0) or ,FALSE).
    Exact match (last arg=0 or FALSE) is required for priority-to-quantity lookups.
    """
    if not formula_str or not isinstance(formula_str, str):
        return False
    f = normalize_formula(formula_str)
    if not f.startswith('=VLOOKUP('):
        return False
    return f.endswith(',0)') or f.endswith(',FALSE)')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the expected sheet exists
    if 'Inventory' not in wb.sheetnames:
        print(f"CRITICAL: 'Inventory' sheet not found. Sheets found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # -------------------------------------------------------------------------
    # Component 1: Column C has VLOOKUP approximate-match formulas (0.5 pts)
    # The task requires =VLOOKUP(B<row>,$D$2:$F$5,2,1) pattern — approximate match
    # to look up stock % against threshold table and return priority level.
    # This FAILS on initial (column C is empty) and PASSES on golden.
    # -------------------------------------------------------------------------
    try:
        col_c_approx_count = sum(
            1 for row_num in PRODUCT_ROWS
            if is_vlookup_approximate(ws.cell(row=row_num, column=3).value)
        )
        col_c_any_vlookup = sum(
            1 for row_num in PRODUCT_ROWS
            if isinstance(ws.cell(row=row_num, column=3).value, str)
            and 'VLOOKUP' in ws.cell(row=row_num, column=3).value.upper()
        )

        if col_c_approx_count >= TOTAL_ROWS:
            print(f"PASS: Component 1 — Column C has VLOOKUP approximate-match in all "
                  f"{col_c_approx_count}/{TOTAL_ROWS} rows. (0.5 pts)")
            total_score += 0.5
        elif col_c_approx_count > 0:
            # Partial: some rows have VLOOKUP approximate match
            frac = col_c_approx_count / TOTAL_ROWS
            pts = round(0.5 * frac, 2)
            print(f"PARTIAL: Component 1 — Column C has VLOOKUP approximate-match in "
                  f"{col_c_approx_count}/{TOTAL_ROWS} rows. Partial: {pts} pts")
            if pts > 0:
                total_score += pts
        elif col_c_any_vlookup > 0:
            print(f"FAIL: Component 1 — Column C has {col_c_any_vlookup} VLOOKUP formula(s) "
                  f"but none use approximate match (last arg=1 or TRUE). "
                  f"Expected approximate match for range-based priority lookup.")
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas found in Column C (rows 2-16). "
                  f"Column C should contain VLOOKUP priority level assignments.")
    except Exception as e:
        print(f"ERROR: Component 1 (Column C VLOOKUP check) — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Column G has VLOOKUP exact-match formulas (0.3 pts)
    # The task requires =VLOOKUP(C<row>,$E$2:$F$5,2,0) pattern — exact match
    # to look up the priority label and return the reorder quantity.
    # This FAILS on initial (column G is empty) and PASSES on golden.
    # -------------------------------------------------------------------------
    try:
        col_g_exact_count = sum(
            1 for row_num in PRODUCT_ROWS
            if is_vlookup_exact(ws.cell(row=row_num, column=7).value)
        )
        col_g_any_vlookup = sum(
            1 for row_num in PRODUCT_ROWS
            if isinstance(ws.cell(row=row_num, column=7).value, str)
            and 'VLOOKUP' in ws.cell(row=row_num, column=7).value.upper()
        )

        if col_g_exact_count >= TOTAL_ROWS:
            print(f"PASS: Component 2 — Column G has VLOOKUP exact-match in all "
                  f"{col_g_exact_count}/{TOTAL_ROWS} rows. (0.3 pts)")
            total_score += 0.3
        elif col_g_exact_count > 0:
            frac = col_g_exact_count / TOTAL_ROWS
            pts = round(0.3 * frac, 2)
            print(f"PARTIAL: Component 2 — Column G has VLOOKUP exact-match in "
                  f"{col_g_exact_count}/{TOTAL_ROWS} rows. Partial: {pts} pts")
            if pts > 0:
                total_score += pts
        elif col_g_any_vlookup > 0:
            print(f"FAIL: Component 2 — Column G has {col_g_any_vlookup} VLOOKUP formula(s) "
                  f"but none use exact match (last arg=0 or FALSE). "
                  f"Expected exact match to look up priority->quantity.")
        else:
            print(f"FAIL: Component 2 — No VLOOKUP formulas found in Column G (rows 2-16). "
                  f"Column G should contain VLOOKUP reorder quantity assignments.")
    except Exception as e:
        print(f"ERROR: Component 2 (Column G VLOOKUP check) — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Both columns have complete coverage — all 15 rows filled (0.2 pts)
    # Verifies the VLOOKUP formulas span the entire dataset (rows 2-16), not just a subset.
    # This FAILS on initial (both columns empty) and PASSES on golden (all 15 rows).
    # -------------------------------------------------------------------------
    try:
        col_c_any_count = sum(
            1 for row_num in PRODUCT_ROWS
            if isinstance(ws.cell(row=row_num, column=3).value, str)
            and 'VLOOKUP' in ws.cell(row=row_num, column=3).value.upper()
        )
        col_g_any_count = sum(
            1 for row_num in PRODUCT_ROWS
            if isinstance(ws.cell(row=row_num, column=7).value, str)
            and 'VLOOKUP' in ws.cell(row=row_num, column=7).value.upper()
        )

        if col_c_any_count >= TOTAL_ROWS and col_g_any_count >= TOTAL_ROWS:
            print(f"PASS: Component 3 — Both columns fully covered: "
                  f"C={col_c_any_count}/{TOTAL_ROWS}, G={col_g_any_count}/{TOTAL_ROWS}. (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Incomplete coverage: "
                  f"Column C = {col_c_any_count}/{TOTAL_ROWS} rows, "
                  f"Column G = {col_g_any_count}/{TOTAL_ROWS} rows. "
                  f"Expected all {TOTAL_ROWS} rows in both columns.")
    except Exception as e:
        print(f"ERROR: Component 3 (full coverage check) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
