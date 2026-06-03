"""
Initial Setup: Inventory reorder priority using VLOOKUP approximate match
Task ID: osworld_calc_vlookup_grade_lookup_010
Domain: libreoffice_calc

Creates a spreadsheet with:
- Column A: Product names
- Column B: Stock % of max capacity
- Column C: EMPTY (agent will add VLOOKUP for priority level)
- Column D: Threshold % (priority lookup table)
- Column E: Priority level label
- Column F: Reorder quantity
- Column G: EMPTY (agent will add second VLOOKUP for reorder quantity)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_grade_lookup_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # --- Header row ---
    headers_left = ['Product', 'Stock %']
    # Column C header: 'Priority' (but no data — agent fills this)
    headers_right = ['Threshold %', 'Priority Level', 'Reorder Qty']
    # Column G header: 'Reorder Qty (Assigned)' (but no data — agent fills this)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    table_fill = PatternFill(start_color="FFE2EFDA", end_color="FFE2EFDA", fill_type="solid")

    # Set column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 22

    # Inventory data headers (row 1)
    ws.cell(row=1, column=1, value='Product').font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=2, value='Stock %').font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=3, value='Priority').font = header_font
    ws.cell(row=1, column=3).fill = header_fill
    # Gap at column D-F for the lookup table header
    ws.cell(row=1, column=4, value='Threshold %').font = header_font
    ws.cell(row=1, column=4).fill = table_fill
    ws.cell(row=1, column=5, value='Priority Level').font = header_font
    ws.cell(row=1, column=5).fill = table_fill
    ws.cell(row=1, column=6, value='Reorder Qty').font = header_font
    ws.cell(row=1, column=6).fill = table_fill
    ws.cell(row=1, column=7, value='Reorder Qty (Assigned)').font = header_font
    ws.cell(row=1, column=7).fill = header_fill

    # --- Inventory data (columns A and B only, column C empty) ---
    # 15 products with varied stock levels to make the VLOOKUP meaningful
    inventory_data = [
        ('Organic Quinoa 500g',        0.05),  # 5% -> Critical
        ('Premium Coffee Beans 1kg',   0.12),  # 12% -> Critical
        ('Almond Milk 1L',             0.22),  # 22% -> High
        ('Whole Grain Crackers 300g',  0.35),  # 35% -> High
        ('Extra Virgin Olive Oil 750ml', 0.48), # 48% -> High
        ('Greek Yogurt 500g',          0.51),  # 51% -> Medium
        ('Dark Chocolate 200g',        0.63),  # 63% -> Medium
        ('Himalayan Salt 1kg',         0.70),  # 70% -> Medium
        ('Wild Caught Tuna 185g',      0.82),  # 82% -> Low
        ('Brown Rice 2kg',             0.88),  # 88% -> Low
        ('Lentil Soup 400g',           0.03),  # 3%  -> Critical
        ('Chia Seeds 500g',            0.25),  # 25% -> High
        ('Coconut Water 330ml',        0.55),  # 55% -> Medium
        ('Rolled Oats 1kg',            0.91),  # 91% -> Low
        ('Dried Cranberries 250g',     0.15),  # 15% -> Critical
    ]

    for r, (product, stock_pct) in enumerate(inventory_data, 2):
        ws.cell(row=r, column=1, value=product)
        ws.cell(row=r, column=2, value=stock_pct)
        ws.cell(row=r, column=2).number_format = '0%'
        # Column C intentionally left empty — agent will add VLOOKUP here
        # Column G intentionally left empty — agent will add second VLOOKUP here

    # --- Priority lookup table (columns D, E, F) ---
    # Approximate match VLOOKUP requires sorted ascending thresholds
    priority_table = [
        (0,    'Critical', 500),
        (0.20, 'High',     250),
        (0.50, 'Medium',   100),
        (0.80, 'Low',       0),
    ]

    for r, (threshold, priority, reorder_qty) in enumerate(priority_table, 2):
        ws.cell(row=r, column=4, value=threshold)
        ws.cell(row=r, column=4).number_format = '0%'
        ws.cell(row=r, column=5, value=priority)
        ws.cell(row=r, column=6, value=reorder_qty)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
