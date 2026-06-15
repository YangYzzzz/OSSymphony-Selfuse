"""
Initial Setup: IFERROR with VLOOKUP for course credits lookup
Task ID: calc_lf_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Courses'

    # Credits lookup table (A1:B4)
    ws['A1'] = 'Code'
    ws['B1'] = 'Credits'
    ws['A2'] = 'CS101'
    ws['B2'] = 3
    ws['A3'] = 'MA201'
    ws['B3'] = 4
    ws['A4'] = 'PH102'
    ws['B4'] = 3

    # Student's course list (D1:E4)
    ws['D1'] = 'My Courses'
    ws['E1'] = 'Credits'
    ws['D2'] = 'CS101'
    ws['D3'] = 'EN300'
    ws['D4'] = 'MA201'
    # E2:E4 intentionally left empty — task is to fill these with IFERROR(VLOOKUP(...))

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
