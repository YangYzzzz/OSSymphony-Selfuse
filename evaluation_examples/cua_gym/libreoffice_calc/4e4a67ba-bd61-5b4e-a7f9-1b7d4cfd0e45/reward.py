"""
Reward Script: IFERROR with VLOOKUP to display 0 for missing course codes
Task ID: calc_lf_021
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): E2 contains IFERROR(VLOOKUP(...)) referencing D2
  Component 2 (0.30): E3 contains IFERROR(VLOOKUP(...)) referencing D3
  Component 3 (0.30): E4 contains IFERROR(VLOOKUP(...)) referencing D4
  Component 4 (0.05): All formulas use absolute range $A$2:$B$4 and fallback 0
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_021'


def normalize_formula(val):
    """Strip spaces and uppercase for comparison."""
    if not isinstance(val, str):
        return ""
    return val.upper().replace(" ", "")


def check_iferror_vlookup(formula_str, expected_lookup_cell):
    """
    Verify formula matches pattern: =IFERROR(VLOOKUP(<cell>,<range>,2,0|FALSE),0)
    Returns (has_iferror, has_vlookup, has_correct_lookup_cell, has_absolute_range, has_fallback_zero)
    """
    norm = normalize_formula(formula_str)
    if not norm.startswith("="):
        return False, False, False, False, False

    has_iferror = "IFERROR(" in norm
    has_vlookup = "VLOOKUP(" in norm
    has_correct_lookup_cell = expected_lookup_cell.upper() in norm
    # Check for absolute range $A$2:$B$4
    has_absolute_range = "$A$2:$B$4" in norm
    # Check fallback value is 0 — the second arg of IFERROR
    # Pattern: ,0) at the end of the IFERROR
    has_fallback_zero = norm.endswith(",0)")

    return has_iferror, has_vlookup, has_correct_lookup_cell, has_absolute_range, has_fallback_zero


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Courses' not in wb.sheetnames:
        print("CRITICAL: 'Courses' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Courses']

    # Track absolute-range consistency for Component 4
    # Initialize as passing; each component sets to False if its check fails
    abs_pass_count = 0
    fallback_pass_count = 0
    formula_count = 0

    # Component 1: E2 contains IFERROR(VLOOKUP(D2,...)) (0.35 points)
    try:
        e2_val = ws['E2'].value
        iferror, vlookup, correct_cell, abs_range, fallback = check_iferror_vlookup(e2_val, "D2")
        if iferror and vlookup and correct_cell:
            print(f"PASS: Component 1 — E2 has IFERROR+VLOOKUP referencing D2 (0.35 pts): {e2_val}")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — E2 expected IFERROR(VLOOKUP(D2,...),0), found: {e2_val}")
            print(f"  iferror={iferror}, vlookup={vlookup}, correct_cell={correct_cell}")
        if abs_range:
            abs_pass_count += 1
        if fallback:
            fallback_pass_count += 1
        formula_count += 1
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: E3 contains IFERROR(VLOOKUP(D3,...)) (0.30 points)
    try:
        e3_val = ws['E3'].value
        iferror, vlookup, correct_cell, abs_range, fallback = check_iferror_vlookup(e3_val, "D3")
        if iferror and vlookup and correct_cell:
            print(f"PASS: Component 2 — E3 has IFERROR+VLOOKUP referencing D3 (0.30 pts): {e3_val}")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — E3 expected IFERROR(VLOOKUP(D3,...),0), found: {e3_val}")
            print(f"  iferror={iferror}, vlookup={vlookup}, correct_cell={correct_cell}")
        if abs_range:
            abs_pass_count += 1
        if fallback:
            fallback_pass_count += 1
        formula_count += 1
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: E4 contains IFERROR(VLOOKUP(D4,...)) (0.30 points)
    try:
        e4_val = ws['E4'].value
        iferror, vlookup, correct_cell, abs_range, fallback = check_iferror_vlookup(e4_val, "D4")
        if iferror and vlookup and correct_cell:
            print(f"PASS: Component 3 — E4 has IFERROR+VLOOKUP referencing D4 (0.30 pts): {e4_val}")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — E4 expected IFERROR(VLOOKUP(D4,...),0), found: {e4_val}")
            print(f"  iferror={iferror}, vlookup={vlookup}, correct_cell={correct_cell}")
        if abs_range:
            abs_pass_count += 1
        if fallback:
            fallback_pass_count += 1
        formula_count += 1
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: All formulas use absolute range $A$2:$B$4 and fallback 0 (0.05 points)
    try:
        if formula_count == 3 and abs_pass_count == 3 and fallback_pass_count == 3:
            print(f"PASS: Component 4 — All formulas use $A$2:$B$4 and fallback 0 (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — abs={abs_pass_count}/3, fallback={fallback_pass_count}/3")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
