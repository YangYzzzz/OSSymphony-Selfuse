"""
Initial Setup: Multi-app task - Download climate paper PDF and document citations
Task ID: osworld_multi_apps_pdf_download_cite_005
Domain: libreoffice_calc (multi-app: Calc + Chrome)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_download_cite_005'
DESKTOP = f'{WORKDIR}/Desktop'
OUTPUT = f'{DESKTOP}/climate_papers.xlsx'


def launch_gui(command: str, delay_sec: float = 1.5):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Climate Papers"

    # --- Header row ---
    headers = ['Title', 'DOI URL', 'Year']
    header_font = Font(bold=True, size=12, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Climate papers data ---
    # Row 1 (data row 2): The IPCC 1.5C report - this is what the agent needs to download
    # Rows 2-5 (data rows 3-6): Related papers, one of which cites the IPCC report
    data = [
        [
            'Global warming of 1.5\u00b0C',
            'https://www.ipcc.ch/sr15/',
            2018
        ],
        [
            'Limiting Global Warming to 1.5\u00b0C: Implications for Carbon Budgets',
            'https://doi.org/10.1038/s41558-018-0091-3',
            2018
        ],
        [
            'Climate change impacts under 1.5\u00b0C and 2\u00b0C of global warming: a focus on the IPCC SR1.5 report',
            'https://doi.org/10.1007/s10584-019-02445-0',
            2019
        ],
        [
            'Net-zero emissions pathways: implications of the IPCC Special Report on Global Warming of 1.5\u00b0C',
            'https://doi.org/10.1038/s41558-018-0336-1',
            2018
        ],
        [
            'Renewable energy and climate policy alignment in the post-Paris era',
            'https://doi.org/10.1016/j.enpol.2020.111271',
            2020
        ],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Set column widths for readability
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 10

    # Freeze the header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    # Open Chrome first (agent will use it to download the PDF)
    launch_gui('google-chrome --new-window "https://www.ipcc.ch/sr15/"', delay_sec=3.0)

    # Open LibreOffice Calc with the climate papers spreadsheet
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
