"""
Initial Setup: Add a page header with company name and date to a quarterly report
Task ID: calc_gsi_024
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Q3 Financial Summary ---
    ws = wb.active
    ws.title = 'Q3 Financial Summary'

    # Header row styling
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    headers = ['Month', 'Revenue', 'COGS', 'Gross Profit', 'Operating Expenses',
               'Net Income', 'Margin %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Monthly data rows for Acme Corp Q3 report
    data = [
        ['July 2025',      487500, 195000, 292500, 134200, 158300, 0.3247],
        ['August 2025',    512300, 204920, 307380, 138750, 168630, 0.3292],
        ['September 2025', 498700, 199480, 299220, 136400, 162820, 0.3265],
    ]

    currency_fmt = '$#,##0'
    pct_fmt = '0.00%'

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        for c in range(1, 6):
            cell = ws.cell(row=r, column=c + 1, value=row_data[c])
            cell.number_format = currency_fmt
        pct_cell = ws.cell(row=r, column=7, value=row_data[6])
        pct_cell.number_format = pct_fmt

    # Totals row
    total_row = 5
    ws.cell(row=total_row, column=1, value='Q3 Total')
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=total_row, column=col)
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col)
        cell.value = f'=SUM({col_letter}2:{col_letter}4)'
        cell.number_format = currency_fmt
        cell.font = Font(bold=True)
    avg_cell = ws.cell(row=total_row, column=7)
    avg_cell.value = '=AVERAGE(G2:G4)'
    avg_cell.number_format = pct_fmt
    avg_cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 16
    ws.column_dimensions['G'].width = 12

    # --- Sheet 2: Department Breakdown ---
    ws2 = wb.create_sheet('Department Breakdown')

    dept_headers = ['Department', 'Headcount', 'Q3 Budget', 'Q3 Actual', 'Variance']
    dept_header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    dept_header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    for col, h in enumerate(dept_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = dept_header_font
        cell.fill = dept_header_fill
        cell.alignment = Alignment(horizontal='center')

    dept_data = [
        ['Engineering',     42,  620000,  598400,  21600],
        ['Sales',           28,  485000,  512300, -27300],
        ['Marketing',       15,  320000,  298750,  21250],
        ['Operations',      22,  410000,  405200,   4800],
        ['HR',               8,  180000,  172500,   7500],
        ['Finance',         12,  265000,  258900,   6100],
        ['Customer Support', 18,  295000,  301400,  -6400],
        ['Legal',            6,  195000,  188700,   6300],
        ['R&D',             20,  540000,  527600,  12400],
        ['Executive',        5,  350000,  345000,   5000],
    ]

    for r, row_data in enumerate(dept_data, 2):
        ws2.cell(row=r, column=1, value=row_data[0])
        ws2.cell(row=r, column=2, value=row_data[1])
        for c in [3, 4, 5]:
            cell = ws2.cell(row=r, column=c, value=row_data[c - 1])
            cell.number_format = '$#,##0'

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 12
    for col_letter in ['C', 'D', 'E']:
        ws2.column_dimensions[col_letter].width = 15

    # --- Sheet 3: Key Metrics ---
    ws3 = wb.create_sheet('Key Metrics')

    metrics_headers = ['Metric', 'Target', 'Actual', 'Status']
    for col, h in enumerate(metrics_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    metrics_data = [
        ['Customer Acquisition Cost',  '$125',  '$118', 'On Track'],
        ['Monthly Active Users',       '45,000', '47,230', 'Exceeded'],
        ['Churn Rate',                 '3.5%',  '3.2%', 'On Track'],
        ['NPS Score',                  '72',    '75', 'Exceeded'],
        ['Average Order Value',        '$89',   '$92', 'Exceeded'],
        ['Support Ticket Resolution',  '24h',   '18h', 'Exceeded'],
        ['Employee Satisfaction',      '4.2',   '4.5', 'Exceeded'],
        ['Revenue per Employee',       '$28K',  '$27.5K', 'Below Target'],
        ['Pipeline Coverage',          '3.0x',  '3.4x', 'On Track'],
        ['Customer Lifetime Value',    '$2,400', '$2,580', 'Exceeded'],
    ]

    for r, row_data in enumerate(metrics_data, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 30
    for col_letter in ['B', 'C', 'D']:
        ws3.column_dimensions[col_letter].width = 15

    # NO page header configured - that is the task for the agent
    # ws.oddHeader is left at default (empty)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
