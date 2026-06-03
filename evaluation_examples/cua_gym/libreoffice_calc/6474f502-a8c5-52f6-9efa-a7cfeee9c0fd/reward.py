"""
Reward Script: Apply double underline to bottom border of A30:F30 (grand total row)
Task ID: calc_gfl_090
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 6 cells A30:F30 have bottom border style == 'double'
  Component 2 (0.3): No other borders (top/left/right) added to A30:F30
  Component 3 (0.2): Adjacent rows (29, 31) have no new borders introduced
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_090'


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Income' sheet must exist
    if 'Income' not in wb.sheetnames:
        print(f"CRITICAL: 'Income' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Income']

    # Component 1: All 6 cells A30:F30 have bottom border style == 'double' (0.5 points)
    try:
        double_count = 0
        for col in range(1, 7):
            cell = ws.cell(row=30, column=col)
            bottom_style = cell.border.bottom.style if cell.border.bottom else None
            if bottom_style == 'double':
                double_count += 1
            else:
                print(f"FAIL: Component 1 — {cell.coordinate} bottom border style is '{bottom_style}', expected 'double'")

        if double_count == 6:
            print(f"PASS: Component 1 — All 6 cells A30:F30 have double bottom border (0.5 pts)")
            total_score += 0.5
        elif double_count > 0:
            # Partial credit: proportional to how many cells have the correct border
            partial = round(0.5 * (double_count / 6), 2)
            print(f"PARTIAL: Component 1 — {double_count}/6 cells have double bottom border ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in A30:F30 have double bottom border")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: No other borders (top/left/right) on A30:F30 (0.3 points)
    # The task specifies bottom border only, no other borders should be added
    try:
        extra_borders_found = False
        for col in range(1, 7):
            cell = ws.cell(row=30, column=col)
            b = cell.border
            top_style = b.top.style if b.top else None
            left_style = b.left.style if b.left else None
            right_style = b.right.style if b.right else None
            if top_style or left_style or right_style:
                print(f"FAIL: Component 2 — {cell.coordinate} has extra borders: top={top_style}, left={left_style}, right={right_style}")
                extra_borders_found = True

        if not extra_borders_found:
            # Only award if Component 1 found at least some double borders
            # (otherwise this check passes trivially on initial state)
            bottom_styles = []
            for col in range(1, 7):
                cell = ws.cell(row=30, column=col)
                bs = cell.border.bottom.style if cell.border.bottom else None
                bottom_styles.append(bs)
            has_any_double = any(s == 'double' for s in bottom_styles)

            if has_any_double:
                print(f"PASS: Component 2 — No extra borders (top/left/right) on A30:F30 (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — No double bottom borders found, so no-extra-borders check is moot")
        else:
            print(f"FAIL: Component 2 — Extra borders found on row 30")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Adjacent rows (29, 31) have no new borders (0.2 points)
    # This ensures the task only modified row 30
    try:
        adjacent_borders_found = False
        for r in [29, 31]:
            for col in range(1, 7):
                cell = ws.cell(row=r, column=col)
                if cell.value is None and r == 31:
                    continue  # row 31 may not exist
                b = cell.border
                for side_name, side in [('bottom', b.bottom), ('top', b.top), ('left', b.left), ('right', b.right)]:
                    style = side.style if side else None
                    if style is not None:
                        print(f"FAIL: Component 3 — Row {r} {cell.coordinate} has {side_name}={style}")
                        adjacent_borders_found = True

        if not adjacent_borders_found:
            # Only award if Component 1 found at least some double borders
            bottom_styles = []
            for col in range(1, 7):
                cell = ws.cell(row=30, column=col)
                bs = cell.border.bottom.style if cell.border.bottom else None
                bottom_styles.append(bs)
            has_any_double = any(s == 'double' for s in bottom_styles)

            if has_any_double:
                print(f"PASS: Component 3 — Adjacent rows (29, 31) have no borders (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — No double bottom borders on row 30, so adjacent-row check is moot")
        else:
            print(f"FAIL: Component 3 — Unexpected borders found on adjacent rows")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
