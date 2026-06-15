"""
Initial Setup: Apply double underline to bottom border of grand total row
Task ID: calc_gfl_090
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Income"

    # --- Headers (Row 1) ---
    headers = ["Category", "Q1", "Q2", "Q3", "Q4", "Annual Total"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # --- Income line items (Rows 2-29) ---
    income_items = [
        ["Product Sales - Electronics", 125400, 138200, 142800, 156300, None],
        ["Product Sales - Furniture", 87600, 92100, 88400, 103500, None],
        ["Product Sales - Office Supplies", 34200, 31800, 36500, 38900, None],
        ["Service Revenue - Consulting", 67800, 72400, 69500, 74200, None],
        ["Service Revenue - Maintenance", 23400, 24100, 25600, 26800, None],
        ["Service Revenue - Training", 15600, 18200, 16900, 19400, None],
        ["Subscription Revenue - Monthly", 45200, 47800, 51200, 54600, None],
        ["Subscription Revenue - Annual", 112000, 112000, 118000, 118000, None],
        ["Licensing Fees - Software", 28900, 31200, 33400, 35100, None],
        ["Licensing Fees - Patents", 8500, 8500, 9200, 9200, None],
        ["Rental Income - Equipment", 12300, 12300, 12300, 12300, None],
        ["Rental Income - Office Space", 36000, 36000, 36000, 36000, None],
        ["Commission Income", 19800, 22400, 21300, 24600, None],
        ["Interest Income", 4200, 4500, 4800, 5100, None],
        ["Dividend Income", 7800, 7800, 8200, 8200, None],
        ["Advertising Revenue", 31500, 34200, 37800, 41200, None],
        ["Affiliate Revenue", 8900, 10200, 11500, 12800, None],
        ["Government Grants", 25000, 0, 25000, 0, None],
        ["Insurance Reimbursements", 3200, 1800, 4500, 2100, None],
        ["Late Payment Fees", 1200, 1500, 1100, 1800, None],
        ["Foreign Exchange Gains", 5600, 3200, 7800, 4100, None],
        ["Asset Disposal Revenue", 0, 15000, 0, 8500, None],
        ["Warranty Revenue", 6700, 7100, 7400, 7900, None],
        ["Franchise Fees", 18000, 18000, 19500, 19500, None],
        ["Sponsorship Income", 10000, 12500, 10000, 15000, None],
        ["Data Monetization Revenue", 22100, 24800, 27500, 30200, None],
        ["Returns and Allowances", -8400, -9200, -7800, -10500, None],
        ["Miscellaneous Income", 2400, 3100, 2800, 3600, None],
    ]

    # Write data rows 2-29
    for r, row_data in enumerate(income_items, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        for c in range(1, 5):
            cell = ws.cell(row=r, column=c + 1, value=row_data[c])
            cell.number_format = '#,##0'
        # Annual Total formula in column F
        ws.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})')
        ws.cell(row=r, column=6).number_format = '#,##0'

    # --- Grand Total Row (Row 30) ---
    ws.cell(row=30, column=1, value="GRAND TOTAL")
    ws.cell(row=30, column=1).font = Font(name="Calibri", size=11, bold=True)
    for col in range(2, 7):
        cell = ws.cell(row=30, column=col)
        cell.value = f'=SUM({chr(64+col)}2:{chr(64+col)}29)'
        cell.number_format = '#,##0'
        cell.font = Font(name="Calibri", size=11, bold=True)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 35
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 15

    # NO borders on row 30 - the task is to add them

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
