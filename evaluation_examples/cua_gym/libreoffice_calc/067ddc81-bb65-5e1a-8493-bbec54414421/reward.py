"""
Reward Script: Rental Property Investment Analyzer
Task ID: calc_wf_065
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Key metrics formulas in B24-B30 (PMT, cash flow, NOI, CoC, cap rate)
  Component 2 (0.15): Key metrics number formatting (currency / percentage)
  Component 3 (0.25): 10-year projection table with formulas (B38:I47)
  Component 4 (0.15): 10-year projection number formatting (currency / percentage)
  Component 5 (0.15): Two line charts present (cash flow trend + equity buildup)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_065'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Analysis' sheet exists
    if 'Analysis' not in wb.sheetnames:
        print("FAIL: 'Analysis' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Analysis']

    # =========================================================================
    # Component 1: Key Metrics Formulas (0.30 points)
    # B24=PMT formula, B25-B30 = calculated metrics.
    # These cells are empty (None) in initial_env.
    # =========================================================================
    try:
        # Count how many key metric cells are filled
        metrics_refs = ['B24', 'B25', 'B26', 'B27', 'B28', 'B29', 'B30']
        filled_count = sum(1 for ref in metrics_refs if ws[ref].value is not None)

        # Check for PMT formula in B24
        b24_val = ws['B24'].value
        has_pmt = isinstance(b24_val, str) and 'PMT' in b24_val.upper()

        if filled_count == 7 and has_pmt:
            total_score += 0.30
            print(f"PASS: Component 1 — All 7 key metric cells filled, PMT formula present (0.30 pts)")
        elif filled_count >= 5 and has_pmt:
            total_score += 0.20
            print(f"PARTIAL: Component 1 — {filled_count}/7 cells filled, PMT found (0.20 pts)")
        elif filled_count >= 3:
            total_score += 0.10
            print(f"PARTIAL: Component 1 — {filled_count}/7 cells filled (0.10 pts)")
        else:
            print(f"FAIL: Component 1 — Only {filled_count}/7 key metric cells filled, PMT found={has_pmt}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Key Metrics Number Formatting (0.15 points)
    # B24-B28 should have currency formatting ($), B29-B30 should have % formatting
    # In initial_env these cells are None with 'General' format.
    # =========================================================================
    try:
        currency_cells = ['B24', 'B25', 'B26', 'B27', 'B28']
        pct_cells = ['B29', 'B30']

        # Count cells that have values AND correct formatting
        has_values = sum(1 for ref in currency_cells + pct_cells if ws[ref].value is not None)
        currency_ok = sum(1 for ref in currency_cells
                         if ws[ref].value is not None
                         and ('$' in (ws[ref].number_format or '') or '#,##0' in (ws[ref].number_format or '')))
        pct_ok = sum(1 for ref in pct_cells
                     if ws[ref].value is not None
                     and '%' in (ws[ref].number_format or ''))

        # Only award points if cells have values (differentiates from initial)
        if has_values >= 5 and currency_ok >= 4 and pct_ok >= 1:
            total_score += 0.15
            print(f"PASS: Component 2 — Currency formatting on {currency_ok}/5 cells, % on {pct_ok}/2 cells (0.15 pts)")
        elif has_values >= 3 and (currency_ok >= 2 or pct_ok >= 1):
            total_score += 0.07
            print(f"PARTIAL: Component 2 — Some formatting correct (0.07 pts)")
        else:
            print(f"FAIL: Component 2 — Insufficient formatting (has_values={has_values}, currency={currency_ok}, pct={pct_ok})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: 10-Year Projection Table Formulas (0.25 points)
    # Rows 38-47 (years 1-10), columns B-I should have formulas/values.
    # In initial_env, these cells are all empty.
    # =========================================================================
    try:
        filled_rows = 0
        for row_num in range(38, 48):
            filled_in_row = sum(1 for col_num in range(2, 10)
                                if ws.cell(row=row_num, column=col_num).value is not None)
            if filled_in_row >= 6:
                filled_rows += 1

        if filled_rows == 10:
            total_score += 0.25
            print(f"PASS: Component 3 — All 10 projection rows have data (0.25 pts)")
        elif filled_rows >= 7:
            total_score += 0.15
            print(f"PARTIAL: Component 3 — {filled_rows}/10 projection rows filled (0.15 pts)")
        elif filled_rows >= 3:
            total_score += 0.08
            print(f"PARTIAL: Component 3 — {filled_rows}/10 projection rows filled (0.08 pts)")
        else:
            print(f"FAIL: Component 3 — Only {filled_rows}/10 projection rows have data")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: 10-Year Projection Number Formatting (0.15 points)
    # Currency formatting on property value/rent/expenses/NOI/cash flow/equity columns,
    # percentage formatting on cash-on-cash return column.
    # Only score if projection data exists (differentiates from initial).
    # =========================================================================
    try:
        row38_has_data = ws.cell(row=38, column=2).value is not None

        if row38_has_data:
            # Check currency formatting on columns B-H (2-8) for row 38
            currency_fmt_count = sum(1 for col_num in range(2, 9)
                                     if '$' in (ws.cell(row=38, column=col_num).number_format or '')
                                     or '#,##0' in (ws.cell(row=38, column=col_num).number_format or ''))

            # Check percentage formatting on column I (9) for any projection row
            pct_fmt_found = any('%' in (ws.cell(row=r, column=9).number_format or '')
                                for r in range(38, 48))

            if currency_fmt_count >= 5 and pct_fmt_found:
                total_score += 0.15
                print(f"PASS: Component 4 — Projection formatting correct: {currency_fmt_count} currency cols, pct present (0.15 pts)")
            elif currency_fmt_count >= 3 or pct_fmt_found:
                total_score += 0.07
                print(f"PARTIAL: Component 4 — Some projection formatting (0.07 pts)")
            else:
                print(f"FAIL: Component 4 — Projection formatting insufficient (currency={currency_fmt_count}, pct={pct_fmt_found})")
        else:
            print(f"FAIL: Component 4 — No projection data in row 38, cannot score formatting")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Line Charts for Cash Flow and Equity (0.15 points)
    # Two line charts expected: cash flow trend and equity buildup.
    # Initial_env has 0 charts.
    # =========================================================================
    try:
        charts = ws._charts
        chart_count = len(charts)
        line_chart_count = sum(1 for c in charts if 'line' in c.tagname.lower())

        if chart_count >= 2 and line_chart_count >= 2:
            total_score += 0.15
            print(f"PASS: Component 5 — {line_chart_count} line charts found (0.15 pts)")
        elif chart_count >= 2 and line_chart_count >= 1:
            total_score += 0.10
            print(f"PARTIAL: Component 5 — {line_chart_count} line chart(s) + {chart_count} total charts (0.10 pts)")
        elif chart_count >= 2:
            total_score += 0.07
            print(f"PARTIAL: Component 5 — {chart_count} charts found but not line type (0.07 pts)")
        elif chart_count == 1:
            total_score += 0.07
            print(f"PARTIAL: Component 5 — Only 1 chart found, expected 2 (0.07 pts)")
        else:
            print(f"FAIL: Component 5 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
