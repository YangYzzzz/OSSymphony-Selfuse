"""
Initial Setup: Rental Property Investment Analyzer
Task ID: calc_wf_065
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Analysis ---
    ws = wb.active
    ws.title = 'Analysis'

    # Styling
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    section_font = Font(name='Calibri', size=11, bold=True, color='2F5496')
    label_font = Font(name='Calibri', size=11)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    # Column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 5
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 18
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 18

    # ===== TITLE =====
    ws.merge_cells('A1:B1')
    ws['A1'] = 'Rental Property Investment Analyzer'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5496')
    ws['A1'].alignment = Alignment(horizontal='left')

    # ===== PROPERTY & MORTGAGE INPUTS =====
    row = 3
    ws.cell(row=row, column=1, value='PROPERTY & MORTGAGE DETAILS').font = section_font
    row = 4
    inputs = [
        ('Purchase Price', 300000),
        ('Down Payment (%)', 0.20),
        ('Down Payment ($)', 60000),
        ('Loan Amount', 240000),
        ('Annual Interest Rate', 0.07),
        ('Loan Term (Years)', 30),
    ]
    for label, val in inputs:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).border = thin_border
        c = ws.cell(row=row, column=2, value=val)
        c.border = thin_border
        c.alignment = Alignment(horizontal='right')
        row += 1

    # Format the input cells
    ws['B4'].number_format = '$#,##0'
    ws['B5'].number_format = '0%'
    ws['B6'].number_format = '$#,##0'
    ws['B7'].number_format = '$#,##0'
    ws['B8'].number_format = '0.00%'
    ws['B9'].number_format = '0'

    # ===== RENTAL INCOME =====
    row = 11
    ws.cell(row=row, column=1, value='RENTAL INCOME').font = section_font
    row = 12
    ws.cell(row=row, column=1, value='Monthly Rent').font = label_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=2200).border = thin_border
    ws['B12'].number_format = '$#,##0'
    ws['B12'].alignment = Alignment(horizontal='right')

    ws.cell(row=13, column=1, value='Annual Rental Income').font = label_font
    ws.cell(row=13, column=1).border = thin_border
    ws.cell(row=13, column=2, value=26400).border = thin_border
    ws['B13'].number_format = '$#,##0'
    ws['B13'].alignment = Alignment(horizontal='right')

    # ===== ANNUAL OPERATING EXPENSES =====
    row = 15
    ws.cell(row=row, column=1, value='ANNUAL OPERATING EXPENSES').font = section_font
    row = 16
    expenses = [
        ('Property Taxes', 3600),
        ('Insurance', 1200),
        ('Maintenance (10% of rent)', 2640),
        ('Vacancy Reserve (5% of rent)', 1320),
        ('Property Management (8% of rent)', 2112),
        ('Total Annual Expenses', 10872),
    ]
    for label, val in expenses:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).border = thin_border
        c = ws.cell(row=row, column=2, value=val)
        c.border = thin_border
        c.number_format = '$#,##0'
        c.alignment = Alignment(horizontal='right')
        row += 1

    # Bold the total row
    ws.cell(row=21, column=1).font = Font(name='Calibri', size=11, bold=True)
    ws.cell(row=21, column=2).font = Font(name='Calibri', size=11, bold=True)

    # ===== KEY METRICS (labels only, NO formulas) =====
    row = 23
    ws.cell(row=row, column=1, value='KEY METRICS').font = section_font
    row = 24
    metrics = [
        'Monthly Mortgage Payment',
        'Monthly Operating Expenses',
        'Monthly Cash Flow',
        'Annual Cash Flow',
        'Net Operating Income (NOI)',
        'Cash-on-Cash Return',
        'Cap Rate',
    ]
    for label in metrics:
        ws.cell(row=row, column=1, value=label).font = label_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2).border = thin_border
        row += 1

    # ===== ASSUMPTIONS =====
    row = 32
    ws.cell(row=row, column=1, value='PROJECTION ASSUMPTIONS').font = section_font
    row = 33
    ws.cell(row=row, column=1, value='Annual Appreciation Rate').font = label_font
    ws.cell(row=row, column=1).border = thin_border
    ws.cell(row=row, column=2, value=0.03).border = thin_border
    ws['B33'].number_format = '0%'
    ws['B33'].alignment = Alignment(horizontal='right')

    ws.cell(row=34, column=1, value='Annual Rent Increase Rate').font = label_font
    ws.cell(row=34, column=1).border = thin_border
    ws.cell(row=34, column=2, value=0.02).border = thin_border
    ws['B34'].number_format = '0%'
    ws['B34'].alignment = Alignment(horizontal='right')

    # ===== 10-YEAR PROJECTION TABLE (structure only, NO computed values) =====
    row = 36
    ws.cell(row=row, column=1, value='10-YEAR INVESTMENT PROJECTION').font = section_font
    row = 37
    proj_headers = [
        'Year', 'Property Value', 'Annual Rent', 'Annual Expenses',
        'NOI', 'Monthly Cash Flow', 'Annual Cash Flow',
        'Total Equity', 'Cash-on-Cash Return'
    ]
    for col_idx, h in enumerate(proj_headers, 1):
        c = ws.cell(row=row, column=col_idx, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = thin_border

    # Year labels only
    for yr in range(1, 11):
        ws.cell(row=37 + yr, column=1, value=yr).border = thin_border
        ws.cell(row=37 + yr, column=1).alignment = Alignment(horizontal='center')
        # Add borders to empty cells
        for col_idx in range(2, 10):
            ws.cell(row=37 + yr, column=col_idx).border = thin_border

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
