"""
Reward Script: DFG Emmy Noether Program Statistics Table
Task ID: osworld_multi_apps_pdf_stats_table_014
Domain: libreoffice_calc
Scoring:
  - Component 1: Correct headers (Year, Applications, Approvals, Approval Rate (%))  0.25 pts
  - Component 2: Correct 5 data rows for years 2018-2022                             0.25 pts
  - Component 3: Correct Applications and Approvals values for all 5 years           0.25 pts
  - Component 4: Correct Approval Rate values (to 2 decimal places) + format         0.25 pts
  Total: 1.0
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_stats_table_014'
FILE_PATH = f'{WORKDIR}/DFG_Emmy_rates.xlsx'

# Expected data extracted from PDFs (ground truth from golden env)
EXPECTED_DATA = {
    2018: {'applications': 2341, 'approvals': 312, 'rate': 13.33},
    2019: {'applications': 2489, 'approvals': 334, 'rate': 13.42},
    2020: {'applications': 2256, 'approvals': 289, 'rate': 12.81},
    2021: {'applications': 2512, 'approvals': 348, 'rate': 13.85},
    2022: {'applications': 2678, 'approvals': 361, 'rate': 13.48},
}

EXPECTED_HEADERS = ['Year', 'Applications', 'Approvals', 'Approval Rate (%)']

# Gate: File must exist before attempting any import of openpyxl
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
    raise SystemExit(0)

# Now import openpyxl (only needed if file exists)
import openpyxl


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Gate: File must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Correct headers (0.25 points)
    # Task requires: 'Year', 'Applications', 'Approvals', 'Approval Rate (%)'
    try:
        headers = []
        for col in range(1, 5):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val).strip() if val is not None else '')

        # Check all four required headers are present
        headers_match = (len(headers) >= 4 and all(
            headers[i].lower() == EXPECTED_HEADERS[i].lower()
            for i in range(4)
        ))
        if headers_match:
            print(f"PASS: Component 1 — Headers correct: {headers[:4]} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Expected headers {EXPECTED_HEADERS}, found {headers}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Correct 5 data rows for years 2018-2022 (0.25 points)
    # Build year->row map and check all 5 years are present
    year_row_map = {}
    try:
        for row in range(2, ws.max_row + 1):
            year_val = ws.cell(row=row, column=1).value
            if year_val is not None:
                try:
                    year_row_map[int(year_val)] = row
                except (ValueError, TypeError):
                    pass

        expected_years = {2018, 2019, 2020, 2021, 2022}
        found_years = set(year_row_map.keys())
        years_present = expected_years.issubset(found_years)

        if years_present:
            print(f"PASS: Component 2 — All 5 years (2018-2022) present in data (0.25 pts)")
            total_score += 0.25
        else:
            missing = expected_years - found_years
            print(f"FAIL: Component 2 — Missing years: {missing}; found years: {found_years}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct Applications and Approvals values for all 5 years (0.25 points)
    try:
        apps_approvals_correct = 0
        apps_approvals_total = len(EXPECTED_DATA)

        for year, expected in EXPECTED_DATA.items():
            if year not in year_row_map:
                print(f"  FAIL row check: year {year} not found in worksheet")
                continue
            row = year_row_map[year]
            apps_val = ws.cell(row=row, column=2).value
            approv_val = ws.cell(row=row, column=3).value

            apps_ok = (apps_val is not None and int(apps_val) == expected['applications'])
            approv_ok = (approv_val is not None and int(approv_val) == expected['approvals'])

            if apps_ok and approv_ok:
                apps_approvals_correct += 1
            else:
                print(f"  FAIL: Year {year} — Applications: expected {expected['applications']} found {apps_val}, "
                      f"Approvals: expected {expected['approvals']} found {approv_val}")

        if apps_approvals_correct == apps_approvals_total:
            print(f"PASS: Component 3 — All Applications and Approvals values correct ({apps_approvals_correct}/5) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Only {apps_approvals_correct}/{apps_approvals_total} rows have correct Applications/Approvals values")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Correct Approval Rate values and formatted to 2 decimal places (0.25 points)
    # Rate must be within 0.01 tolerance and cell number_format must indicate 2 decimal places
    try:
        rate_correct = 0
        rate_total = len(EXPECTED_DATA)
        format_correct = 0

        for year, expected in EXPECTED_DATA.items():
            if year not in year_row_map:
                continue
            row = year_row_map[year]
            rate_cell = ws.cell(row=row, column=4)
            rate_val = rate_cell.value

            # Check value correctness (within 0.01 tolerance)
            val_ok = False
            if rate_val is not None:
                try:
                    diff = abs(float(rate_val) - expected['rate'])
                    val_ok = (diff <= 0.01)
                except (ValueError, TypeError):
                    pass

            # Check number format for 2 decimal places
            fmt = rate_cell.number_format
            fmt_ok = (fmt is not None and '0.00' in fmt)

            if val_ok:
                rate_correct += 1
            else:
                print(f"  FAIL: Year {year} Approval Rate: expected {expected['rate']}, found {rate_val}")

            if fmt_ok:
                format_correct += 1
            else:
                print(f"  FAIL: Year {year} rate format: expected '0.00' style, found '{fmt}'")

        # Award full points if rates are correct AND formatted
        if rate_correct == rate_total and format_correct == rate_total:
            print(f"PASS: Component 4 — All Approval Rates correct ({rate_correct}/5) and formatted to 2 dp (0.25 pts)")
            total_score += 0.25
        elif rate_correct == rate_total:
            # Values correct but formatting not explicitly set to 0.00
            print(f"PASS (partial): Component 4 — Rates correct ({rate_correct}/5) but format not 0.00 "
                  f"({format_correct}/5 formatted correctly). Awarding 0.15 pts")
            if rate_correct == rate_total:
                total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Only {rate_correct}/{rate_total} approval rates correct; "
                  f"{format_correct}/{rate_total} properly formatted")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


verify_task(FILE_PATH)
