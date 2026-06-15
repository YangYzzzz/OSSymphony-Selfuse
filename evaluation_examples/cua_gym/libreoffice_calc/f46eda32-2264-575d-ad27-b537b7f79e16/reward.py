"""
Reward Script: Set font color of column headers in row 1 to dark blue (#1F4E79) and make them bold.
Task ID: calc_fmt_font_color_blue_header_007
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 7 header cells A1:G1 have bold=True
  Component 2 (0.5): All 7 header cells A1:G1 have font color FF1F4E79 (dark blue #1F4E79)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_font_color_blue_header_007'
SHEET_NAME = 'Customer Database'
HEADER_COLS = 7  # Columns A through G (1-7)
EXPECTED_FONT_COLOR = 'FF1F4E79'  # ARGB for #1F4E79 (dark blue)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Set the font color of the column headers in row 1 to dark blue (#1F4E79)
    and make them bold.
    Headers are in A1:G1 of the 'Customer Database' sheet.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: All header cells A1:G1 must have bold=True (0.5 points)
    # This FAILS on initial file (all bold=False) and PASSES on golden file (all bold=True)
    try:
        bold_results = []
        for col in range(1, HEADER_COLS + 1):
            cell = ws.cell(row=1, column=col)
            is_bold = cell.font.bold == True
            bold_results.append((cell.coordinate, is_bold))

        all_bold = all(result for _, result in bold_results)
        bold_fails = [coord for coord, result in bold_results if not result]

        if all_bold:
            print(f"PASS: Component 1 — All {HEADER_COLS} header cells (A1:G1) are bold (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Not all header cells are bold. Non-bold cells: {bold_fails}")
    except Exception as e:
        print(f"ERROR: Component 1 (bold check) — {e}")

    # Component 2: All header cells A1:G1 must have font color = FF1F4E79 (0.5 points)
    # This FAILS on initial file (all 00000000 / black) and PASSES on golden file (all FF1F4E79)
    try:
        color_results = []
        for col in range(1, HEADER_COLS + 1):
            cell = ws.cell(row=1, column=col)
            try:
                actual_color = cell.font.color.rgb
            except Exception:
                actual_color = None
            matches = (actual_color == EXPECTED_FONT_COLOR)
            color_results.append((cell.coordinate, actual_color, matches))

        all_correct_color = all(matches for _, _, matches in color_results)
        color_fails = [(coord, actual) for coord, actual, matches in color_results if not matches]

        if all_correct_color:
            print(f"PASS: Component 2 — All {HEADER_COLS} header cells (A1:G1) have font color {EXPECTED_FONT_COLOR} (dark blue) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — Not all header cells have the expected font color {EXPECTED_FONT_COLOR}.")
            for coord, actual in color_fails:
                print(f"  {coord}: expected {EXPECTED_FONT_COLOR}, found {actual}")
    except Exception as e:
        print(f"ERROR: Component 2 (font color check) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
