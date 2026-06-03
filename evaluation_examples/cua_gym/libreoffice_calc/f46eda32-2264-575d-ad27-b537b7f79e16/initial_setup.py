"""
Initial Setup: Customer Database spreadsheet with unformatted headers
Task ID: calc_fmt_font_color_blue_header_007
Domain: libreoffice_calc

Creates a spreadsheet with one sheet 'Customer Database' containing
149 rows of customer data (rows 2-150) with plain black unformatted headers in row 1.
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_font_color_blue_header_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Customer Database'

    # Row 1: Headers — default black font, NO bold, NO color
    headers = ['Customer ID', 'Name', 'Email', 'Phone', 'City', 'Account Value', 'Join Date']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        # Explicitly set default (non-bold, black) font to be clear
        cell.font = Font(bold=False, color='000000')

    # Customer data rows 2-150 (149 records)
    customers = [
        ('C001', 'Sarah Chen', 'sarah.chen@email.com', '(415) 555-0101', 'San Francisco', 12500.00, '2021-03-15'),
        ('C002', 'Marcus Johnson', 'marcus.j@gmail.com', '(312) 555-0182', 'Chicago', 8750.50, '2020-11-02'),
        ('C003', 'Elena Rodriguez', 'erodriguez@company.com', '(305) 555-0247', 'Miami', 23400.00, '2019-07-22'),
        ('C004', 'David Kim', 'dkim@techcorp.net', '(206) 555-0318', 'Seattle', 15600.75, '2022-01-10'),
        ('C005', 'Priya Patel', 'priya.patel@biz.com', '(408) 555-0459', 'San Jose', 31200.00, '2018-09-05'),
        ('C006', 'James O\'Brien', 'jobrien@outlook.com', '(617) 555-0563', 'Boston', 9800.25, '2021-06-18'),
        ('C007', 'Aisha Williams', 'a.williams@webmail.com', '(404) 555-0624', 'Atlanta', 18900.50, '2020-04-30'),
        ('C008', 'Robert Tanaka', 'rtanaka@enterprise.org', '(213) 555-0735', 'Los Angeles', 42100.00, '2017-12-01'),
        ('C009', 'Fatima Hassan', 'fhassan@corp.net', '(713) 555-0846', 'Houston', 7350.00, '2023-02-14'),
        ('C010', 'Christopher Lee', 'clee@domain.com', '(602) 555-0957', 'Phoenix', 27800.25, '2019-10-08'),
        ('C011', 'Natalia Ivanova', 'n.ivanova@mail.ru', '(916) 555-0168', 'Sacramento', 11400.00, '2022-05-20'),
        ('C012', 'Kevin Brown', 'kbrown@startup.io', '(503) 555-0279', 'Portland', 19200.75, '2021-08-11'),
        ('C013', 'Mei Lin', 'mlin@globalcorp.com', '(702) 555-0380', 'Las Vegas', 35600.00, '2018-03-27'),
        ('C014', 'Thomas Anderson', 't.anderson@neo.net', '(615) 555-0491', 'Nashville', 8100.50, '2023-07-03'),
        ('C015', 'Yuki Nakamura', 'ynakamura@japan.co', '(808) 555-0502', 'Honolulu', 22300.00, '2020-01-15'),
        ('C016', 'Angela Davis', 'adavis@nonprofit.org', '(510) 555-0613', 'Oakland', 14700.25, '2021-11-28'),
        ('C017', 'Luis Fernandez', 'lfernandez@latam.com', '(305) 555-0724', 'Miami', 28900.00, '2019-04-09'),
        ('C018', 'Hannah Schmidt', 'hschmidt@german.de', '(312) 555-0835', 'Chicago', 16500.75, '2022-09-14'),
        ('C019', 'Mohammed Al-Rashid', 'malrashid@gulf.ae', '(971) 555-0946', 'Dubai', 48200.00, '2017-06-21'),
        ('C020', 'Sophia Rossi', 's.rossi@italy.it', '(415) 555-0157', 'San Francisco', 10300.50, '2023-01-07'),
        ('C021', 'Benjamin Clark', 'bclark@financial.com', '(212) 555-0268', 'New York', 67500.00, '2016-08-30'),
        ('C022', 'Nina Kowalski', 'nkowalski@poland.pl', '(206) 555-0379', 'Seattle', 13800.25, '2022-03-22'),
        ('C023', 'Carlos Mendoza', 'cmendoza@mexico.mx', '(713) 555-0480', 'Houston', 21400.00, '2020-07-16'),
        ('C024', 'Rachel Green', 'rgreen@friends.tv', '(212) 555-0591', 'New York', 9600.75, '2021-10-05'),
        ('C025', 'Ahmed Khalil', 'akhalil@cairo.eg', '(602) 555-0602', 'Phoenix', 33700.00, '2018-12-19'),
        ('C026', 'Isabelle Dubois', 'idubois@france.fr', '(617) 555-0713', 'Boston', 17200.50, '2022-06-08'),
        ('C027', 'William Zhang', 'wzhang@pacific.com', '(408) 555-0824', 'San Jose', 25800.00, '2019-02-25'),
        ('C028', 'Olga Petrova', 'oprova@russia.ru', '(503) 555-0935', 'Portland', 12100.25, '2023-05-11'),
        ('C029', 'Derek Thompson', 'd.thompson@corp.us', '(916) 555-0146', 'Sacramento', 39400.00, '2018-09-02'),
        ('C030', 'Amara Okafor', 'aokafor@nigeria.ng', '(404) 555-0257', 'Atlanta', 8900.75, '2021-04-24'),
        ('C031', 'Jason Wu', 'jwu@techgiant.com', '(415) 555-0368', 'San Francisco', 54300.00, '2017-01-17'),
        ('C032', 'Valentina Cruz', 'vcruz@spain.es', '(213) 555-0479', 'Los Angeles', 16800.50, '2022-10-31'),
        ('C033', 'Patrick Murphy', 'pmurphy@ireland.ie', '(617) 555-0580', 'Boston', 22900.00, '2020-08-13'),
        ('C034', 'Soo-Yeon Park', 'sypark@korea.kr', '(312) 555-0691', 'Chicago', 31500.25, '2019-05-06'),
        ('C035', 'Nicole Lambert', 'nlambert@canada.ca', '(702) 555-0702', 'Las Vegas', 14400.00, '2021-12-20'),
        ('C036', 'Omar Abdullah', 'oabdullah@saudi.sa', '(713) 555-0813', 'Houston', 62100.75, '2016-03-14'),
        ('C037', 'Chloe Martin', 'cmartin@paris.fr', '(503) 555-0924', 'Portland', 11700.00, '2023-04-02'),
        ('C038', 'Eric Nielsen', 'enielsen@denmark.dk', '(808) 555-0135', 'Honolulu', 28400.50, '2019-11-27'),
        ('C039', 'Lakshmi Gupta', 'lgupta@india.in', '(408) 555-0246', 'San Jose', 19800.00, '2021-07-08'),
        ('C040', 'Michael Torres', 'mtorres@business.net', '(602) 555-0357', 'Phoenix', 37200.25, '2018-06-15'),
        ('C041', 'Zara Ahmed', 'zahmed@london.uk', '(212) 555-0468', 'New York', 10600.00, '2022-12-03'),
        ('C042', 'Daniel Reyes', 'dreyes@colombia.co', '(310) 555-0579', 'Los Angeles', 24500.75, '2020-03-19'),
        ('C043', 'Emma Wilson', 'ewilson@australia.au', '(206) 555-0680', 'Seattle', 16100.00, '2021-09-30'),
        ('C044', 'Victor Popescu', 'vpopescu@romania.ro', '(916) 555-0791', 'Sacramento', 42800.50, '2017-07-23'),
        ('C045', 'Lily Chang', 'lchang@hongkong.hk', '(415) 555-0802', 'San Francisco', 29300.00, '2019-08-10'),
        ('C046', 'Nathan Brooks', 'nbrooks@midwest.com', '(312) 555-0913', 'Chicago', 13200.25, '2023-06-26'),
        ('C047', 'Anastasia Morozova', 'amorozova@moscow.ru', '(503) 555-0124', 'Portland', 21900.00, '2020-10-14'),
        ('C048', 'Ibrahim Al-Farsi', 'ialfarsi@oman.om', '(713) 555-0235', 'Houston', 55600.75, '2016-11-07'),
        ('C049', 'Diana Flores', 'dflores@mexico.mx', '(808) 555-0346', 'Honolulu', 8400.00, '2022-07-21'),
        ('C050', 'Andrew Peterson', 'apeterson@nordic.no', '(617) 555-0457', 'Boston', 34100.50, '2018-04-04'),
        ('C051', 'Yumi Watanabe', 'ywatanabe@tokyo.jp', '(408) 555-0568', 'San Jose', 17700.00, '2021-05-17'),
        ('C052', 'Brandon Foster', 'bfoster@midwest.us', '(602) 555-0679', 'Phoenix', 26400.25, '2019-12-31'),
        ('C053', 'Rosa Jimenez', 'rjimenez@texas.com', '(713) 555-0780', 'Houston', 11200.00, '2023-03-16'),
        ('C054', 'Felix Mueller', 'fmueller@berlin.de', '(212) 555-0891', 'New York', 48700.75, '2017-09-29'),
        ('C055', 'Tanya Koroleva', 'tkoroleva@kiev.ua', '(310) 555-0102', 'Los Angeles', 20100.00, '2020-06-22'),
        ('C056', 'Mark Hoffman', 'mhoffman@finance.com', '(206) 555-0213', 'Seattle', 36800.50, '2018-01-12'),
        ('C057', 'Sun Li', 'sunli@shanghai.cn', '(415) 555-0324', 'San Francisco', 23500.00, '2021-02-28'),
        ('C058', 'Michelle Larson', 'mlarson@midwest.net', '(312) 555-0435', 'Chicago', 15900.25, '2022-11-15'),
        ('C059', 'Antonio Gonzalez', 'agonzalez@madrid.es', '(503) 555-0546', 'Portland', 41200.00, '2017-04-18'),
        ('C060', 'Hina Tanaka', 'htanaka@osaka.jp', '(808) 555-0657', 'Honolulu', 9100.75, '2023-08-01'),
        ('C061', 'George Papadopoulos', 'gpapadopoulos@athens.gr', '(617) 555-0768', 'Boston', 28700.00, '2019-03-12'),
        ('C062', 'Alicia Hernandez', 'ahernandez@cali.com', '(916) 555-0879', 'Sacramento', 14300.50, '2021-06-25'),
        ('C063', 'Ravi Sharma', 'rsharma@mumbai.in', '(408) 555-0980', 'San Jose', 32600.00, '2018-10-08'),
        ('C064', 'Megan Taylor', 'mtaylor@usa.com', '(602) 555-0191', 'Phoenix', 10800.25, '2022-04-19'),
        ('C065', 'Santiago Vargas', 'svargas@bogota.co', '(213) 555-0202', 'Los Angeles', 47500.00, '2016-07-05'),
        ('C066', 'Ingrid Sorensen', 'isorensen@oslo.no', '(206) 555-0313', 'Seattle', 19600.75, '2020-09-28'),
        ('C067', 'Terry Walsh', 'twalsh@dublin.ie', '(712) 555-0424', 'Chicago', 25200.00, '2021-01-11'),
        ('C068', 'Xinwei Gao', 'xgao@beijing.cn', '(415) 555-0535', 'San Francisco', 38300.50, '2018-07-24'),
        ('C069', 'Adele Fontaine', 'afontaine@lyon.fr', '(617) 555-0646', 'Boston', 12700.00, '2023-09-14'),
        ('C070', 'Marcus Webb', 'mwebb@corp.co', '(916) 555-0757', 'Sacramento', 30400.25, '2019-06-07'),
        ('C071', 'Chioma Eze', 'ceze@lagos.ng', '(713) 555-0868', 'Houston', 8600.00, '2022-08-30'),
        ('C072', 'Stefan Novak', 'snovak@prague.cz', '(503) 555-0979', 'Portland', 44900.75, '2016-12-23'),
        ('C073', 'Jennifer Nguyen', 'jnguyen@viet.vn', '(808) 555-0180', 'Honolulu', 17400.00, '2020-05-16'),
        ('C074', 'Adam Blackwell', 'ablackwell@uk.com', '(212) 555-0291', 'New York', 27100.50, '2021-03-09'),
        ('C075', 'Leila Nazari', 'lnazari@tehran.ir', '(408) 555-0302', 'San Jose', 20800.00, '2019-01-21'),
        ('C076', 'Samuel Osei', 'sosei@accra.gh', '(602) 555-0413', 'Phoenix', 35900.25, '2018-05-14'),
        ('C077', 'Vera Lindqvist', 'vlindqvist@stockholm.se', '(310) 555-0524', 'Los Angeles', 11500.00, '2023-11-27'),
        ('C078', 'Peter Kozlowski', 'pkozlowski@warsaw.pl', '(206) 555-0635', 'Seattle', 43600.75, '2017-02-10'),
        ('C079', 'Nadia Moreau', 'nmoreau@bordeaux.fr', '(916) 555-0746', 'Sacramento', 22100.00, '2020-12-04'),
        ('C080', 'Calvin Ho', 'cho@taipei.tw', '(415) 555-0857', 'San Francisco', 29800.50, '2021-08-17'),
        ('C081', 'Zainab Malik', 'zmalik@karachi.pk', '(312) 555-0968', 'Chicago', 14100.00, '2019-09-30'),
        ('C082', 'Luca Ferrari', 'lferrari@milan.it', '(617) 555-0179', 'Boston', 51200.25, '2016-04-13'),
        ('C083', 'Ayasha Redcloud', 'aredcloud@tribal.us', '(503) 555-0280', 'Portland', 9300.00, '2022-02-06'),
        ('C084', 'Hiroshi Yamamoto', 'hyamamoto@kyoto.jp', '(808) 555-0391', 'Honolulu', 37500.75, '2018-08-19'),
        ('C085', 'Tiffany Brooks', 'tbrooks@south.us', '(713) 555-0402', 'Houston', 16200.00, '2020-11-12'),
        ('C086', 'Andrei Volkov', 'avolkov@spb.ru', '(212) 555-0513', 'New York', 24700.50, '2021-04-25'),
        ('C087', 'Grace Okonkwo', 'gokonkwo@abuja.ng', '(408) 555-0624', 'San Jose', 33100.00, '2019-07-18'),
        ('C088', 'Pablo Vega', 'pvega@lima.pe', '(602) 555-0735', 'Phoenix', 10100.25, '2023-10-01'),
        ('C089', 'Miriam Goldstein', 'mgoldstein@tel.il', '(310) 555-0846', 'Los Angeles', 46400.00, '2017-03-24'),
        ('C090', 'Francis Oduya', 'foduya@nairobi.ke', '(206) 555-0957', 'Seattle', 21300.75, '2020-08-06'),
        ('C091', 'Catalina Ruiz', 'cruiz@bogota.co', '(916) 555-0168', 'Sacramento', 28600.00, '2021-07-29'),
        ('C092', 'Ian MacDonald', 'imacdonald@edinburgh.uk', '(415) 555-0279', 'San Francisco', 15700.50, '2019-11-22'),
        ('C093', 'Sasha Belova', 'sbelova@minsk.by', '(312) 555-0380', 'Chicago', 40300.00, '2018-02-15'),
        ('C094', 'Takeshi Ito', 'tito@sapporo.jp', '(617) 555-0491', 'Boston', 12400.25, '2023-12-08'),
        ('C095', 'Monica Vasquez', 'mvasquez@caracas.ve', '(503) 555-0502', 'Portland', 25500.00, '2020-04-21'),
        ('C096', 'Lawrence Obi', 'lobi@portharcourtng', '(808) 555-0613', 'Honolulu', 34200.75, '2018-11-04'),
        ('C097', 'Bianca Romano', 'bromano@naples.it', '(713) 555-0724', 'Houston', 18600.00, '2021-10-17'),
        ('C098', 'Esteban Jimenez', 'ejimenez@cancun.mx', '(212) 555-0835', 'New York', 27900.50, '2019-04-30'),
        ('C099', 'Keiko Ogawa', 'kogawa@nagoya.jp', '(408) 555-0946', 'San Jose', 11800.00, '2022-06-13'),
        ('C100', 'Ryan O\'Connor', 'roconnor@cork.ie', '(602) 555-0157', 'Phoenix', 39700.25, '2017-08-06'),
        ('C101', 'Amina Diallo', 'adiallo@dakar.sn', '(310) 555-0268', 'Los Angeles', 14900.00, '2020-02-19'),
        ('C102', 'Tobias Berg', 'tberg@hamburg.de', '(206) 555-0379', 'Seattle', 23800.75, '2021-09-02'),
        ('C103', 'Rosa Martinez', 'rmartinez@havana.cu', '(916) 555-0480', 'Sacramento', 31700.00, '2019-06-25'),
        ('C104', 'Kwame Asante', 'kasante@kumasi.gh', '(415) 555-0591', 'San Francisco', 9700.50, '2023-07-09'),
        ('C105', 'Cecile Beaumont', 'cbeaumont@nice.fr', '(312) 555-0602', 'Chicago', 45100.00, '2016-10-22'),
        ('C106', 'Sergei Morozov', 'smorozov@rostov.ru', '(617) 555-0713', 'Boston', 18300.25, '2021-05-05'),
        ('C107', 'Joanna Wisniewski', 'jwisniewski@krakow.pl', '(503) 555-0824', 'Portland', 26900.00, '2019-10-18'),
        ('C108', 'Akira Hayashi', 'ahayashi@fukuoka.jp', '(808) 555-0935', 'Honolulu', 13500.75, '2022-09-01'),
        ('C109', 'Gabriela Santos', 'gsantos@rio.br', '(713) 555-0146', 'Houston', 32800.00, '2018-03-14'),
        ('C110', 'Hassan Benali', 'hbenali@casablanca.ma', '(212) 555-0257', 'New York', 10400.50, '2023-04-27'),
        ('C111', 'Svetlana Orlova', 'sorlova@kazan.ru', '(408) 555-0368', 'San Jose', 50300.00, '2016-01-10'),
        ('C112', 'Jean-Pierre Leblanc', 'jpleblanc@montreal.ca', '(602) 555-0479', 'Phoenix', 21700.25, '2020-07-23'),
        ('C113', 'Nkechi Adeyemi', 'nadeyemi@ibadan.ng', '(310) 555-0580', 'Los Angeles', 16400.00, '2021-11-06'),
        ('C114', 'Christoph Bauer', 'cbauer@vienna.at', '(206) 555-0691', 'Seattle', 38100.75, '2018-08-29'),
        ('C115', 'Marisol Reyes', 'mreyes@santiago.cl', '(916) 555-0702', 'Sacramento', 23200.00, '2019-05-12'),
        ('C116', 'Kweku Mensah', 'kmensah@tema.gh', '(415) 555-0813', 'San Francisco', 29100.50, '2021-02-25'),
        ('C117', 'Petra Horak', 'phorak@brno.cz', '(312) 555-0924', 'Chicago', 11000.00, '2023-01-18'),
        ('C118', 'Abdullah Rashid', 'arashid@amman.jo', '(617) 555-0135', 'Boston', 43900.25, '2017-06-01'),
        ('C119', 'Consuela Aguilar', 'caguilar@guadalajara.mx', '(503) 555-0246', 'Portland', 20500.00, '2020-10-14'),
        ('C120', 'Shiro Matsumoto', 'smatsumoto@kobe.jp', '(808) 555-0357', 'Honolulu', 35400.75, '2018-12-27'),
        ('C121', 'Destiny Walker', 'dwalker@atlanta.us', '(713) 555-0468', 'Houston', 14600.00, '2021-04-10'),
        ('C122', 'Rudolf Novotny', 'rnovotny@bratislava.sk', '(212) 555-0579', 'New York', 27400.50, '2019-08-23'),
        ('C123', 'Oluwaseun Adeyemi', 'oadeyemi@enugu.ng', '(408) 555-0680', 'San Jose', 18100.00, '2022-07-06'),
        ('C124', 'Marion Lebrun', 'mlebrun@toulouse.fr', '(602) 555-0791', 'Phoenix', 32000.25, '2018-04-19'),
        ('C125', 'Bjorn Eriksson', 'beriksson@gothenburg.se', '(310) 555-0802', 'Los Angeles', 10700.00, '2023-05-02'),
        ('C126', 'Siti Rahayu', 'srahayu@jakarta.id', '(206) 555-0913', 'Seattle', 45600.75, '2016-09-25'),
        ('C127', 'Rodrigo Morales', 'rmorales@montevideo.uy', '(916) 555-0124', 'Sacramento', 22800.00, '2020-06-08'),
        ('C128', 'Fikile Dlamini', 'fdlamini@durban.za', '(415) 555-0235', 'San Francisco', 31100.50, '2021-03-21'),
        ('C129', 'Ming-Wei Chen', 'mwchen@taichung.tw', '(312) 555-0346', 'Chicago', 15300.00, '2019-12-04'),
        ('C130', 'Odalys Pacheco', 'opacheco@quito.ec', '(617) 555-0457', 'Boston', 38800.25, '2018-07-17'),
        ('C131', 'Pieter Van Dijk', 'pvandijk@rotterdam.nl', '(503) 555-0568', 'Portland', 12300.00, '2022-10-30'),
        ('C132', 'Yusuf Abubakar', 'yabubakar@kano.ng', '(808) 555-0679', 'Honolulu', 26700.75, '2020-09-12'),
        ('C133', 'Elisa Moretti', 'emoretti@florence.it', '(713) 555-0780', 'Houston', 17900.00, '2021-08-25'),
        ('C134', 'Patrick Osei-Bonsu', 'poseibon@takoradi.gh', '(212) 555-0891', 'New York', 34600.50, '2019-03-08'),
        ('C135', 'Kazuki Fujita', 'kfujita@hiroshima.jp', '(408) 555-0102', 'San Jose', 10200.00, '2023-02-21'),
        ('C136', 'Susannah Price', 'sprice@sydney.au', '(602) 555-0213', 'Phoenix', 47800.25, '2017-01-04'),
        ('C137', 'Nikolai Vasiliev', 'nvasiliev@volgograd.ru', '(310) 555-0324', 'Los Angeles', 21200.00, '2020-05-28'),
        ('C138', 'Adeleke Oladipo', 'aoladipo@abeokuta.ng', '(206) 555-0435', 'Seattle', 28900.75, '2021-06-10'),
        ('C139', 'Camille Rousseau', 'crousseau@strasbourg.fr', '(916) 555-0546', 'Sacramento', 14000.00, '2019-11-24'),
        ('C140', 'Mihail Popescu', 'mpopescu@bucharest.ro', '(415) 555-0657', 'San Francisco', 36300.50, '2018-09-06'),
        ('C141', 'Tunde Okafor', 'tokafor@benin-city.ng', '(312) 555-0768', 'Chicago', 11900.00, '2022-12-19'),
        ('C142', 'Ana Pereira', 'apereira@lisbon.pt', '(617) 555-0879', 'Boston', 44200.25, '2016-06-02'),
        ('C143', 'Ichiro Sasaki', 'isasaki@sendai.jp', '(503) 555-0980', 'Portland', 20700.00, '2020-11-15'),
        ('C144', 'Emmanuel Nwosu', 'enwosu@port-harcourt.ng', '(808) 555-0191', 'Honolulu', 30500.75, '2021-05-28'),
        ('C145', 'Greta Hoffmann', 'ghoffmann@cologne.de', '(713) 555-0202', 'Houston', 13700.00, '2019-02-10'),
        ('C146', 'Alexei Petrov', 'apetrov@novosibirsk.ru', '(212) 555-0313', 'New York', 48600.50, '2016-08-23'),
        ('C147', 'Fatou Coulibaly', 'fcoulibaly@bamako.ml', '(408) 555-0424', 'San Jose', 19300.00, '2020-03-07'),
        ('C148', 'Xavier Dumont', 'xdumont@brussels.be', '(602) 555-0535', 'Phoenix', 37100.25, '2018-11-20'),
        ('C149', 'Josephine Nakagawa', 'jnakagawa@yokohama.jp', '(310) 555-0646', 'Los Angeles', 24400.00, '2021-10-03'),
    ]

    for r, row_data in enumerate(customers, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Customer Database with {len(customers)} customer records')
    print('Headers in row 1: plain black, non-bold (no formatting applied to task elements)')


create_initial()
