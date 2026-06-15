"""
Initial Setup: Create spreadsheet with employee performance scores (no sparklines)
Task ID: calc_gcp_054
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PerformanceScores'

    # --- Headers ---
    headers = ['Employee', 'Month1', 'Month2', 'Month3', 'Month4', 'Month5', 'Trend']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_side = Side(style='thin', color='000000')
    header_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Employee Data (30 rows) ---
    # Design varied trends: some up, some down, some flat, some mixed
    employees = [
        ('Sarah Chen',          78, 80, 83, 86, 90),     # trending up
        ('Marcus Johnson',      92, 90, 88, 85, 82),     # trending down
        ('Priya Patel',         75, 76, 75, 77, 76),     # flat
        ('David Kim',           65, 70, 74, 80, 85),     # strong upward
        ('Elena Rodriguez',     88, 86, 84, 82, 80),     # gradual decline
        ('James O\'Brien',      72, 78, 71, 79, 73),     # fluctuating
        ('Aisha Mohammed',      90, 91, 93, 94, 96),     # steady climb
        ('Robert Fischer',      68, 65, 63, 60, 62),     # declining then slight recovery
        ('Wei Zhang',           82, 85, 88, 91, 94),     # consistent improvement
        ('Maria Santos',        77, 77, 78, 77, 78),     # very flat
        ('Thomas Anderson',     95, 92, 89, 86, 83),     # steady decline
        ('Fatima Al-Rashid',    60, 65, 72, 78, 85),     # strong improvement
        ('Liam Murphy',         84, 82, 86, 83, 87),     # oscillating up
        ('Yuki Tanaka',         71, 75, 79, 83, 87),     # linear increase
        ('Sophie Dubois',       88, 90, 87, 91, 89),     # slight fluctuation high
        ('Carlos Mendez',       76, 73, 70, 67, 64),     # declining
        ('Ingrid Svensson',     80, 82, 84, 86, 88),     # steady increase
        ('Omar Hassan',         69, 74, 68, 75, 70),     # volatile
        ('Rachel Green',        85, 87, 89, 91, 93),     # consistent rise
        ('Nikolai Petrov',      93, 91, 94, 90, 92),     # fluctuating high
        ('Amara Okafor',        62, 68, 73, 79, 84),     # accelerating improvement
        ('Benjamin Wright',     87, 85, 83, 81, 79),     # gradual decline
        ('Mei Lin',             74, 78, 82, 86, 90),     # strong upward
        ('Patrick Sullivan',    81, 80, 79, 78, 77),     # slow decline
        ('Zara Khan',           66, 72, 78, 84, 90),     # linear increase steep
        ('Henrik Johansson',    89, 88, 87, 86, 85),     # very gradual decline
        ('Diana Popescu',       73, 76, 80, 77, 82),     # mostly up with dip
        ('Alejandro Ruiz',      91, 93, 95, 97, 99),     # near-perfect trajectory
        ('Nadia Kowalski',      79, 75, 80, 76, 81),     # sawtooth pattern
        ('Samuel Achebe',       70, 73, 76, 79, 82),     # steady linear rise
    ]

    data_font = Font(name='Calibri', size=11)
    data_align_left = Alignment(horizontal='left', vertical='center')
    data_align_center = Alignment(horizontal='center', vertical='center')
    data_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for r, (name, m1, m2, m3, m4, m5) in enumerate(employees, 2):
        # Employee name
        cell = ws.cell(row=r, column=1, value=name)
        cell.font = data_font
        cell.alignment = data_align_left
        cell.border = data_border

        # Month scores
        for c, score in enumerate([m1, m2, m3, m4, m5], 2):
            cell = ws.cell(row=r, column=c, value=score)
            cell.font = data_font
            cell.alignment = data_align_center
            cell.border = data_border
            cell.number_format = '0'

        # Column G (Trend) - intentionally left EMPTY for the task
        cell = ws.cell(row=r, column=7)
        cell.border = data_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 10
    ws.column_dimensions['G'].width = 18

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
