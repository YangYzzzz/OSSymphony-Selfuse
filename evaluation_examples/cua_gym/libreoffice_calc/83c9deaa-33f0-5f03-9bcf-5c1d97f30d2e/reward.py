"""
Reward Script: Sparkline formulas in column G for employee performance trends
Task ID: calc_gcp_054
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): SPARKLINE formulas exist in G2:G31
  Component 2 (0.3): Each SPARKLINE references the correct row range (Bn:Fn)
  Component 3 (0.2): All 30 employee rows have sparkline formulas (complete coverage)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_054'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the PerformanceScores sheet
    try:
        ws = wb['PerformanceScores']
    except KeyError:
        # Try active sheet as fallback
        ws = wb.active
        if ws is None:
            print("CRITICAL: No accessible sheet found")
            print("REWARD: 0.0")
            return 0.0
        print(f"WARN: 'PerformanceScores' not found, using active sheet '{ws.title}'")

    # Component 1: SPARKLINE formulas exist in G2:G31 (0.5 points)
    # Count how many cells in G2:G31 contain a SPARKLINE formula
    try:
        sparkline_count = 0
        for r in range(2, 32):
            cell_val = ws.cell(row=r, column=7).value
            if cell_val is not None and isinstance(cell_val, str):
                # Check for SPARKLINE function (case-insensitive)
                if 'SPARKLINE' in cell_val.upper() and cell_val.strip().startswith('='):
                    sparkline_count += 1

        if sparkline_count >= 25:
            # At least 25 out of 30 have sparkline formulas
            ratio = sparkline_count / 30.0
            component1_score = 0.5 * ratio
            print(f"PASS: Component 1 - {sparkline_count}/30 cells have SPARKLINE formulas ({component1_score:.3f} pts)")
            total_score += component1_score
        elif sparkline_count > 0:
            # Partial credit: some sparklines exist
            ratio = sparkline_count / 30.0
            component1_score = 0.5 * ratio
            print(f"PARTIAL: Component 1 - {sparkline_count}/30 cells have SPARKLINE formulas ({component1_score:.3f} pts)")
            total_score += component1_score
        else:
            print(f"FAIL: Component 1 - No SPARKLINE formulas found in G2:G31")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Each SPARKLINE references the correct row range Bn:Fn (0.3 points)
    # Verify the formula pattern =SPARKLINE(Bn:Fn) where n is the row number
    try:
        correct_ref_count = 0
        sparkline_with_formula = 0
        for r in range(2, 32):
            cell_val = ws.cell(row=r, column=7).value
            if cell_val is not None and isinstance(cell_val, str) and 'SPARKLINE' in cell_val.upper():
                sparkline_with_formula += 1
                # Expected pattern: =SPARKLINE(B{r}:F{r}) - possibly with extra args
                # Normalize: remove spaces, uppercase
                normalized = cell_val.upper().replace(' ', '')
                expected_range = f'B{r}:F{r}'
                if expected_range in normalized:
                    correct_ref_count += 1

        if sparkline_with_formula > 0:
            ratio = correct_ref_count / 30.0
            component2_score = 0.3 * ratio
            if correct_ref_count >= 25:
                print(f"PASS: Component 2 - {correct_ref_count}/30 sparklines reference correct row ranges ({component2_score:.3f} pts)")
            else:
                print(f"PARTIAL: Component 2 - {correct_ref_count}/30 sparklines reference correct row ranges ({component2_score:.3f} pts)")
            total_score += component2_score
        else:
            print(f"FAIL: Component 2 - No SPARKLINE formulas to check references")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: All 30 employee rows covered (0.2 points)
    # This checks completeness - all rows G2 through G31 must have sparklines
    try:
        all_covered = True
        missing_rows = []
        for r in range(2, 32):
            cell_val = ws.cell(row=r, column=7).value
            if cell_val is None or not isinstance(cell_val, str) or 'SPARKLINE' not in cell_val.upper():
                all_covered = False
                missing_rows.append(r)

        if all_covered:
            print(f"PASS: Component 3 - All 30 rows (G2:G31) have sparkline formulas (0.2 pts)")
            total_score += 0.2
        elif len(missing_rows) <= 5:
            # Minor gaps - partial credit
            covered = 30 - len(missing_rows)
            component3_score = 0.2 * (covered / 30.0)
            print(f"PARTIAL: Component 3 - {covered}/30 rows covered, missing rows: {missing_rows[:10]} ({component3_score:.3f} pts)")
            total_score += component3_score
        else:
            print(f"FAIL: Component 3 - {len(missing_rows)} rows missing sparklines (first 10: {missing_rows[:10]})")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
