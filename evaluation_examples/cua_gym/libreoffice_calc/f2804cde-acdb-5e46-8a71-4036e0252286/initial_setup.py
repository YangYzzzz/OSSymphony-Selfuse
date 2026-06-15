"""
Initial Setup: HR Employee ID Find & Replace
Task ID: calc_hr_employee_id_find_replace_041
Domain: libreoffice_calc

Creates a spreadsheet with 120 employees with EMP-XXXX format IDs.
The task is to replace the 'EMP-' prefix with 'EC-' for all IDs.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_employee_id_find_replace_041'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Roster'

    # --- Header row ---
    headers = ['Employee ID', 'First Name', 'Last Name', 'Department', 'Job Title', 'Hire Date', 'Salary']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')

    # --- Realistic employee data ---
    first_names = [
        'James', 'Mary', 'Robert', 'Patricia', 'John', 'Jennifer', 'Michael', 'Linda',
        'William', 'Barbara', 'David', 'Elizabeth', 'Richard', 'Susan', 'Joseph', 'Jessica',
        'Thomas', 'Sarah', 'Charles', 'Karen', 'Christopher', 'Lisa', 'Daniel', 'Nancy',
        'Matthew', 'Betty', 'Anthony', 'Margaret', 'Mark', 'Sandra', 'Donald', 'Ashley',
        'Steven', 'Dorothy', 'Paul', 'Kimberly', 'Andrew', 'Emily', 'Kenneth', 'Donna',
        'Joshua', 'Michelle', 'Kevin', 'Carol', 'Brian', 'Amanda', 'George', 'Melissa',
        'Timothy', 'Deborah', 'Ronald', 'Stephanie', 'Edward', 'Rebecca', 'Jason', 'Laura',
        'Jeffrey', 'Sharon', 'Ryan', 'Cynthia', 'Jacob', 'Kathleen', 'Gary', 'Amy',
        'Nicholas', 'Angela', 'Eric', 'Shirley', 'Jonathan', 'Anna', 'Stephen', 'Brenda',
        'Larry', 'Pamela', 'Justin', 'Emma', 'Scott', 'Nicole', 'Brandon', 'Helen',
        'Benjamin', 'Samantha', 'Samuel', 'Katherine', 'Nathan', 'Christine', 'Gregory', 'Debra',
        'Frank', 'Rachel', 'Raymond', 'Carolyn', 'Alexander', 'Janet', 'Patrick', 'Catherine',
        'Jack', 'Maria', 'Dennis', 'Heather', 'Jerry', 'Diane', 'Tyler', 'Julie',
        'Aaron', 'Joyce', 'Jose', 'Victoria', 'Henry', 'Kelly', 'Adam', 'Christina',
        'Douglas', 'Ruth', 'Zachary', 'Joan', 'Peter', 'Evelyn', 'Kyle', 'Judith',
        'Walter', 'Andrea', 'Ethan', 'Megan',
    ]

    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
        'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez', 'Wilson', 'Anderson',
        'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin', 'Lee', 'Perez', 'Thompson',
        'White', 'Harris', 'Sanchez', 'Clark', 'Ramirez', 'Lewis', 'Robinson', 'Walker',
        'Young', 'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores',
        'Green', 'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts', 'Gomez', 'Phillips', 'Evans', 'Turner', 'Diaz', 'Parker',
        'Cruz', 'Edwards', 'Collins', 'Reyes', 'Stewart', 'Morris', 'Morales', 'Murphy',
        'Cook', 'Rogers', 'Gutierrez', 'Ortiz', 'Morgan', 'Cooper', 'Peterson', 'Bailey',
        'Reed', 'Kelly', 'Howard', 'Ramos', 'Kim', 'Cox', 'Ward', 'Richardson',
        'Watson', 'Brooks', 'Chavez', 'Wood', 'James', 'Bennett', 'Gray', 'Mendoza',
        'Ruiz', 'Hughes', 'Price', 'Alvarez', 'Castillo', 'Sanders', 'Patel', 'Myers',
        'Long', 'Ross', 'Foster', 'Jimenez', 'Powell', 'Jenkins', 'Perry', 'Russell',
        'Sullivan', 'Bell', 'Coleman', 'Butler', 'Henderson', 'Barnes', 'Gonzales', 'Fisher',
        'Vasquez', 'Simmons', 'Romero', 'Jordan', 'Patterson', 'Alexander', 'Hamilton', 'Graham',
        'Reynolds', 'Griffin', 'Wallace', 'Moreno', 'West', 'Cole', 'Hayes', 'Bryant',
    ]

    departments = [
        'Engineering', 'Marketing', 'Sales', 'Human Resources', 'Finance',
        'Operations', 'Customer Support', 'Legal', 'Product Management', 'IT',
    ]

    job_titles = {
        'Engineering': ['Software Engineer', 'Senior Engineer', 'Lead Engineer', 'QA Engineer', 'DevOps Engineer'],
        'Marketing': ['Marketing Specialist', 'Marketing Manager', 'Content Writer', 'SEO Analyst', 'Brand Manager'],
        'Sales': ['Sales Representative', 'Account Executive', 'Sales Manager', 'Business Developer', 'Sales Analyst'],
        'Human Resources': ['HR Specialist', 'HR Manager', 'Recruiter', 'HR Coordinator', 'Benefits Analyst'],
        'Finance': ['Financial Analyst', 'Accountant', 'Finance Manager', 'Controller', 'Budget Analyst'],
        'Operations': ['Operations Manager', 'Operations Analyst', 'Project Manager', 'Process Engineer', 'Supply Chain Analyst'],
        'Customer Support': ['Support Specialist', 'Support Manager', 'Technical Support', 'Client Success Manager', 'Help Desk Analyst'],
        'Legal': ['Legal Counsel', 'Paralegal', 'Compliance Officer', 'Contract Manager', 'Legal Analyst'],
        'Product Management': ['Product Manager', 'Product Owner', 'Product Analyst', 'UX Designer', 'Product Lead'],
        'IT': ['Systems Administrator', 'Network Engineer', 'IT Manager', 'Security Analyst', 'Database Administrator'],
    }

    # Generate hire dates (various years 2018-2024)
    hire_dates = []
    months = ['2018-03-15', '2018-07-22', '2019-01-08', '2019-04-30', '2019-09-14',
              '2020-02-03', '2020-06-19', '2020-10-27', '2021-01-11', '2021-05-25',
              '2021-08-16', '2022-02-28', '2022-06-07', '2022-11-03', '2023-01-23',
              '2023-04-17', '2023-07-31', '2023-10-09', '2024-01-15', '2024-03-20']
    for i in range(120):
        hire_dates.append(months[i % len(months)])

    # Generate salaries (realistic range by department)
    salary_ranges = {
        'Engineering': (75000, 145000),
        'Marketing': (55000, 105000),
        'Sales': (50000, 120000),
        'Human Resources': (52000, 95000),
        'Finance': (65000, 130000),
        'Operations': (58000, 110000),
        'Customer Support': (42000, 75000),
        'Legal': (80000, 155000),
        'Product Management': (85000, 150000),
        'IT': (70000, 130000),
    }

    # Write 120 employee rows
    import random
    random.seed(42)  # reproducible data

    # Generate unique 4-digit IDs
    id_numbers = random.sample(range(1000, 9999), 120)
    id_numbers.sort()  # sort for readability

    for i in range(120):
        row = i + 2
        emp_id = f'EMP-{id_numbers[i]:04d}'
        first = first_names[i % len(first_names)]
        last = last_names[i % len(last_names)]
        dept = departments[i % len(departments)]
        titles = job_titles[dept]
        title = titles[i % len(titles)]
        hire_date = hire_dates[i]
        sal_min, sal_max = salary_ranges[dept]
        salary = random.randint(sal_min // 1000, sal_max // 1000) * 1000

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=first)
        ws.cell(row=row, column=3, value=last)
        ws.cell(row=row, column=4, value=dept)
        ws.cell(row=row, column=5, value=title)
        ws.cell(row=row, column=6, value=hire_date)
        ws.cell(row=row, column=7, value=salary)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Roster')
    print(f'  Rows: 1 header + 120 employee rows (rows 2-121)')
    print(f'  Employee IDs: EMP-XXXX format in column A')


create_initial()
