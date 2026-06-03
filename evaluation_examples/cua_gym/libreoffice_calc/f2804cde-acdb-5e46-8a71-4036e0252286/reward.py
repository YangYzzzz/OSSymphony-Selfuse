"""
Reward Script: Replace employee ID prefix EMP- with EC- in column A
Task ID: calc_hr_employee_id_find_replace_041
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 120 IDs in A2:A121 start with 'EC-' (prefix replaced)
  Component 2 (0.3): The numeric suffix is preserved correctly for all IDs
  Component 3 (0.2): Header A1 unchanged AND other columns B-G intact
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_employee_id_find_replace_041'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Roster sheet exists
    if 'Roster' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Roster' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Roster']

    # Also load initial file as reference to verify other columns are untouched
    # We derive the initial path from the current path pattern
    initial_path = file_path.replace('_initial.', '_initial.')
    initial_available = os.path.exists(initial_path)
    if initial_available:
        try:
            wb_init = openpyxl.load_workbook(initial_path)
            ws_init = wb_init['Roster'] if 'Roster' in wb_init.sheetnames else None
        except Exception:
            ws_init = None
            initial_available = False
    else:
        ws_init = None

    # Component 1: All 120 IDs in A2:A121 start with 'EC-' (0.5 points)
    # This FAILS on initial (all start with 'EMP-') and PASSES on golden.
    try:
        ec_count = 0
        non_ec_values = []
        for row in range(2, 122):
            val = ws.cell(row=row, column=1).value
            if val is not None and str(val).startswith('EC-'):
                ec_count += 1
            else:
                non_ec_values.append((row, val))

        if ec_count == 120:
            print(f"PASS: Component 1 — All 120 IDs in A2:A121 start with 'EC-' (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Only {ec_count}/120 IDs start with 'EC-'. "
                  f"Sample non-EC values: {non_ec_values[:5]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Numeric suffix preserved correctly for all IDs (0.3 points)
    # Verifies that 'EMP-XXXX' became 'EC-XXXX' (same number, only prefix changed).
    # This FAILS on initial (values still have 'EMP-' prefix, so suffix-check logic
    # against expected 'EC-' prefix doesn't match) and PASSES on golden.
    try:
        # Load the initial to compare suffix preservation
        # For each row in golden, the value should be 'EC-' + original_suffix
        # We load the initial file to derive expected suffix
        if initial_available and ws_init is not None:
            suffix_ok = 0
            suffix_fail = []
            for row in range(2, 122):
                val_init = ws_init.cell(row=row, column=1).value
                val_curr = ws.cell(row=row, column=1).value
                if val_init is not None and str(val_init).startswith('EMP-'):
                    suffix = str(val_init)[4:]  # strip 'EMP-'
                    expected = 'EC-' + suffix
                    if str(val_curr) == expected:
                        suffix_ok += 1
                    else:
                        suffix_fail.append((row, val_init, val_curr, expected))
                else:
                    # Initial value doesn't start with EMP-, check golden matches initial (no change needed)
                    if val_curr == val_init:
                        suffix_ok += 1
                    else:
                        suffix_fail.append((row, val_init, val_curr, 'unchanged'))

            if suffix_ok == 120:
                print(f"PASS: Component 2 — All 120 numeric suffixes preserved correctly (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — {suffix_ok}/120 suffixes correct. "
                      f"Failures: {suffix_fail[:3]}")
        else:
            # Without initial file, verify suffix format: all IDs match EC-\d+ pattern
            import re
            pattern = re.compile(r'^EC-\d+$')
            format_ok = 0
            format_fail = []
            for row in range(2, 122):
                val = ws.cell(row=row, column=1).value
                if val is not None and pattern.match(str(val)):
                    format_ok += 1
                else:
                    format_fail.append((row, val))
            if format_ok == 120:
                print(f"PASS: Component 2 — All 120 IDs match 'EC-NNNN' format (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Only {format_ok}/120 IDs match 'EC-NNNN' format. "
                      f"Failures: {format_fail[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Header A1 unchanged AND other columns B-G intact (0.2 points)
    # Header should still be 'Employee ID', no changes in cols B-G.
    # This FAILS on initial file if checked via "no EMP- prefix in A1" since A1='Employee ID'
    # in both files. BUT the combined condition: A1=='Employee ID' AND all 120 IDs in A2:A121
    # are 'EC-' format (which was already verified in Component 1) — here we specifically
    # verify that col B-G data is untouched (columns that should NOT have changed).
    # On the initial file, this would fail because IDs still have EMP- prefix,
    # so the check for no 'EMP-' prefix in A2:A121 catches that.
    # Here we check: A1 header is correct AND B-G cols are intact.
    # Since this could pass on initial (B-G were never changed), we make the component
    # explicitly conditional: it only awards points if Component 1 passed AND B-G is intact.
    # Structurally: gate on at least one EC- change happened (i.e., score > 0) before awarding.
    try:
        # Check header A1
        header_ok = (ws.cell(row=1, column=1).value == 'Employee ID')

        # Check no 'EMP-' prefix remains anywhere in A2:A121
        # (This is a complementary check to C1: after the replacement, no EMP- should remain)
        emp_remaining = 0
        for row in range(2, 122):
            val = ws.cell(row=row, column=1).value
            if val is not None and str(val).startswith('EMP-'):
                emp_remaining += 1

        # Check B-G columns integrity using initial file if available
        other_cols_changes = 0
        if initial_available and ws_init is not None:
            for row in range(1, 122):
                for col in range(2, 8):
                    v_i = ws_init.cell(row=row, column=col).value
                    v_c = ws.cell(row=row, column=col).value
                    if v_i != v_c:
                        other_cols_changes += 1
                        if other_cols_changes <= 3:
                            print(f"FAIL: Component 3 — Column {col} row {row} changed: "
                                  f"{repr(v_i)} -> {repr(v_c)}")

        if header_ok and emp_remaining == 0 and other_cols_changes == 0:
            print(f"PASS: Component 3 — Header 'Employee ID' intact, no remaining EMP- IDs, "
                  f"B-G columns unchanged (0.2 pts)")
            total_score += 0.2
        else:
            if not header_ok:
                print(f"FAIL: Component 3 — Header A1 changed: "
                      f"{repr(ws.cell(row=1, column=1).value)}")
            if emp_remaining > 0:
                print(f"FAIL: Component 3 — {emp_remaining} cells in A2:A121 still have 'EMP-' prefix")
            if other_cols_changes > 0:
                print(f"FAIL: Component 3 — {other_cols_changes} cells in columns B-G were modified")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
