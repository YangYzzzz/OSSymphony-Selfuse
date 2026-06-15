"""
Initial Setup: Purchase order spreadsheet with partial formula in D2 only.
Task ID: osworld_calc_formula_pattern_concat_015
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Purchase Order ---
    ws = wb.active
    ws.title = "Purchase Order"

    # Headers
    headers = ['Item Name', 'Quantity', 'Unit Price', 'Total Cost', 'Summary']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic purchase order data (10 rows)
    data = [
        ['Mechanical Keyboard',      4,   89.99,  None, None],
        ['USB-C Monitor',            2,  349.00,  None, None],
        ['Wireless Mouse',          10,   34.50,  None, None],
        ['Desk Lamp LED',            6,   27.95,  None, None],
        ['Laptop Stand Aluminum',    3,   55.00,  None, None],
        ['HDMI Cable 2m',           15,    8.75,  None, None],
        ['Noise-Cancelling Headset', 5,  119.90,  None, None],
        ['Webcam HD 1080p',          8,   69.00,  None, None],
        ['Power Strip 6-Outlet',     7,   22.49,  None, None],
        ['External SSD 1TB',         4,  109.95,  None, None],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Item Name
        ws.cell(row=r, column=2, value=row_data[1])   # Quantity
        ws.cell(row=r, column=3, value=row_data[2])   # Unit Price
        # Column D: only D2 has the formula (task requires filling the rest)
        # Column E: all empty (task requires creating these)

    # D2 has the formula as specified in context
    ws['D2'] = '=B2*C2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
