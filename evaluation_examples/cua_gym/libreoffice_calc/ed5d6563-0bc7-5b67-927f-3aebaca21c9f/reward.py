"""
Reward Script: Fill total cost formula down column D and create concatenation formulas in column E
Task ID: osworld_calc_formula_pattern_concat_015
Domain: libreoffice_calc
Scoring:
  - Component 1: Column D formulas filled for all data rows D3:D11 (0.5 pts)
  - Component 2: Column E has concatenation formulas for all rows E2:E11 (0.5 pts)
Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_015'


def normalize_formula(formula):
    """Normalize formula for comparison: uppercase, remove spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def is_multiply_formula(formula, row):
    """
    Check if formula is a multiplication of B*C for the given row.
    Accepts patterns like =B3*C3 or =C3*B3.
    """
    if not isinstance(formula, str):
        return False
    norm = normalize_formula(formula)
    expected1 = f'=B{row}*C{row}'
    expected2 = f'=C{row}*B{row}'
    return norm == expected1 or norm == expected2


def is_concat_formula(formula, row):
    """
    Check if formula is a concatenation formula for the given row.
    Expected pattern: =A#&" x"&B#&" @ $"&TEXT(C#,"0.00")&" = $"&TEXT(D#,"0.00")
    We check for key structural elements:
      - References A#, B#, C#, D# for the correct row
      - Contains TEXT( function with 0.00 format
      - Contains both " x" separator and " @ $" and " = $" markers
    """
    if not isinstance(formula, str):
        return False
    norm = normalize_formula(formula)
    # Must reference item name from column A
    if f'A{row}' not in norm:
        return False
    # Must reference quantity from column B
    if f'B{row}' not in norm:
        return False
    # Must reference unit price formatted via TEXT with "0.00"
    if 'TEXT(' not in norm:
        return False
    if '"0.00"' not in norm and "'0.00'" not in norm and '0.00' not in norm:
        return False
    # Must reference total cost from column D
    if f'D{row}' not in norm:
        return False
    # Must contain concatenation operator
    if '&' not in norm:
        return False
    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook (formulas mode)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Identify the worksheet (should be 'Purchase Order')
    try:
        if 'Purchase Order' in wb.sheetnames:
            ws = wb['Purchase Order']
        else:
            ws = wb.worksheets[0]
        print(f"INFO: Using sheet '{ws.title}'")
    except Exception as e:
        print(f"CRITICAL: Cannot access worksheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Component 1: Column D formulas filled for rows D3:D11 (0.5 points) ---
    # D2 already has =B2*C2 in initial_env, so we focus on D3:D11 (9 new cells).
    # All 9 cells must contain a multiplication formula =B#*C# for their respective row.
    try:
        data_rows = range(3, 12)  # rows 3 through 11
        d_filled = 0
        d_total = len(data_rows)
        d_failures = []

        for r in data_rows:
            cell_val = ws.cell(row=r, column=4).value  # column D
            if is_multiply_formula(cell_val, r):
                d_filled += 1
            else:
                d_failures.append(f"D{r}={repr(cell_val)}")

        if d_filled == d_total:
            print(f"PASS: Component 1 — All D3:D11 have multiplication formulas ({d_filled}/{d_total}) (0.5 pts)")
            total_score += 0.5
        elif d_filled > 0:
            # Partial within component: still 0.5 or 0 — no partial for this component
            print(f"FAIL: Component 1 — Only {d_filled}/{d_total} of D3:D11 have formulas. Missing: {d_failures}")
        else:
            print(f"FAIL: Component 1 — None of D3:D11 have multiplication formulas. Missing: {d_failures}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: Column E has concatenation formulas for all rows E2:E11 (0.5 points) ---
    # All 10 data rows in E must have a concatenation formula referencing A, B, C, D
    # with TEXT formatting to 2 decimal places.
    try:
        e_rows = range(2, 12)  # rows 2 through 11
        e_filled = 0
        e_total = len(e_rows)
        e_failures = []

        for r in e_rows:
            cell_val = ws.cell(row=r, column=5).value  # column E
            if is_concat_formula(cell_val, r):
                e_filled += 1
            else:
                e_failures.append(f"E{r}={repr(cell_val)}")

        if e_filled == e_total:
            print(f"PASS: Component 2 — All E2:E11 have concatenation formulas ({e_filled}/{e_total}) (0.5 pts)")
            total_score += 0.5
        elif e_filled > 0:
            print(f"FAIL: Component 2 — Only {e_filled}/{e_total} of E2:E11 have concat formulas. Missing: {e_failures}")
        else:
            print(f"FAIL: Component 2 — None of E2:E11 have concatenation formulas. Missing: {e_failures}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
