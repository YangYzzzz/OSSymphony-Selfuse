"""
Reward Script: Replace deeply nested IF formula with efficient lookup approach
Task ID: calc_tbl_074
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Lookup helper table exists with correct score-to-category mapping
  Component 2 (0.3): E2 formula replaced with efficient approach (not nested IF)
  Component 3 (0.2): All data rows E2:E21 have formulas applied
  Component 4 (0.2): Formulas produce correct category mappings (via cached values)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_074'

# Expected score-to-category mapping (10 categories)
EXPECTED_CATEGORIES = {
    0: 'Critical',
    10: 'Very Poor',
    20: 'Poor',
    30: 'Below Average',
    40: 'Needs Improvement',
    50: 'Satisfactory',
    60: 'Good',
    70: 'Very Good',
    80: 'Excellent',
    90: 'Outstanding',
}

def score_to_category(score):
    """Map a numeric score to its expected category using the range logic."""
    if score >= 90:
        return 'Outstanding'
    elif score >= 80:
        return 'Excellent'
    elif score >= 70:
        return 'Very Good'
    elif score >= 60:
        return 'Good'
    elif score >= 50:
        return 'Satisfactory'
    elif score >= 40:
        return 'Needs Improvement'
    elif score >= 30:
        return 'Below Average'
    elif score >= 20:
        return 'Poor'
    elif score >= 10:
        return 'Very Poor'
    else:
        return 'Critical'


def is_nested_if(formula_str):
    """Check if a formula is a deeply nested IF (5+ levels)."""
    if not isinstance(formula_str, str):
        return False
    # Count occurrences of IF( in the formula
    count = len(re.findall(r'IF\s*\(', formula_str, re.IGNORECASE))
    return count >= 5


def is_efficient_formula(formula_str):
    """Check if a formula uses an efficient approach (VLOOKUP, IFS, INDEX/MATCH, XLOOKUP, etc.)."""
    if not isinstance(formula_str, str):
        return False
    formula_upper = formula_str.upper()
    efficient_patterns = ['VLOOKUP', 'HLOOKUP', 'IFS', 'INDEX', 'MATCH', 'XLOOKUP', 'CHOOSE', 'SWITCH']
    for pat in efficient_patterns:
        if pat in formula_upper:
            return True
    # Also accept: nested IF with <= 3 levels (simplified)
    if_count = len(re.findall(r'IF\s*\(', formula_str, re.IGNORECASE))
    if 0 < if_count <= 3:
        return True
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Lookup helper table exists with correct mapping (0.3 points)
    # This checks that a new sheet with score-to-category mapping was created.
    # Initial file has only 'Performance Data' sheet; golden adds a 'Lookup' sheet.
    try:
        # Find a sheet that serves as a lookup table (not the main data sheet)
        lookup_sheet = None
        for sn in wb.sheetnames:
            if sn != 'Performance Data':
                ws_candidate = wb[sn]
                # Check if this sheet has score-category mapping data
                # Look for at least 5 rows of numeric -> text mappings
                mapping_count = 0
                for r in range(1, ws_candidate.max_row + 1):
                    val_a = ws_candidate.cell(row=r, column=1).value
                    val_b = ws_candidate.cell(row=r, column=2).value
                    if isinstance(val_a, (int, float)) and isinstance(val_b, str):
                        # Check if the text matches one of our expected categories
                        if val_b.strip() in EXPECTED_CATEGORIES.values():
                            mapping_count += 1
                if mapping_count >= 8:  # at least 8 of 10 categories present
                    lookup_sheet = sn
                    break

        if lookup_sheet is not None:
            print(f"PASS: Component 1 -- Lookup helper table found in sheet '{lookup_sheet}' with valid mappings (0.3 pts)")
            total_score += 0.3
        else:
            # Also accept if the formulas use IFS() without a helper table
            # (task says "such as using a helper lookup table" but also allows other efficient approaches)
            ws_main = wb['Performance Data']
            e2_val = ws_main['E2'].value
            if isinstance(e2_val, str) and 'IFS' in e2_val.upper():
                print(f"PASS: Component 1 -- IFS-based approach used (no lookup table needed) (0.3 pts)")
                total_score += 0.3
            elif isinstance(e2_val, str) and ('SWITCH' in e2_val.upper() or 'CHOOSE' in e2_val.upper()):
                print(f"PASS: Component 1 -- SWITCH/CHOOSE-based approach used (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 -- No lookup helper table found and no IFS/SWITCH approach in E2")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: E2 formula replaced with efficient approach (0.3 points)
    # Initial E2 has deeply nested IF (10 levels). Golden should NOT have nested IF.
    try:
        ws_main = wb['Performance Data']
        e2_val = ws_main['E2'].value

        if e2_val is None:
            print(f"FAIL: Component 2 -- E2 is empty")
        elif not isinstance(e2_val, str) or not e2_val.startswith('='):
            print(f"FAIL: Component 2 -- E2 is not a formula: {repr(e2_val)}")
        elif is_nested_if(e2_val):
            print(f"FAIL: Component 2 -- E2 still has deeply nested IF formula ({len(re.findall(r'IF', e2_val, re.IGNORECASE))} IF levels)")
        elif is_efficient_formula(e2_val):
            print(f"PASS: Component 2 -- E2 uses efficient formula: {e2_val[:80]}... (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 -- E2 formula not recognized as efficient: {e2_val[:80]}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: All data rows E2:E21 have formulas (0.2 points)
    # Initial file only has formula in E2, E3:E21 are empty.
    try:
        ws_main = wb['Performance Data']
        formula_count = 0
        for r in range(2, 22):  # rows 2-21
            val = ws_main.cell(row=r, column=5).value
            if isinstance(val, str) and val.startswith('='):
                formula_count += 1

        if formula_count >= 20:
            print(f"PASS: Component 3 -- All 20 data rows have formulas in column E (0.2 pts)")
            total_score += 0.2
        elif formula_count >= 15:
            partial = round(0.2 * (formula_count / 20), 2)
            print(f"PARTIAL: Component 3 -- {formula_count}/20 rows have formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 -- Only {formula_count}/20 rows have formulas in column E")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Formulas produce correct category results (0.2 points)
    # Load with data_only=True to get cached computed values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['Performance Data']

        # Get D column scores and check E column cached values
        correct_count = 0
        checked = 0
        for r in range(2, 22):
            score_val = ws_data.cell(row=r, column=4).value
            category_val = ws_data.cell(row=r, column=5).value

            if score_val is not None and category_val is not None:
                expected = score_to_category(score_val)
                if str(category_val).strip() == expected:
                    correct_count += 1
                checked += 1

        if checked >= 15 and correct_count >= checked:
            print(f"PASS: Component 4 -- All {correct_count}/{checked} cached values match expected categories (0.2 pts)")
            total_score += 0.2
        elif checked >= 10 and correct_count >= checked * 0.8:
            partial = round(0.2 * (correct_count / checked), 2)
            print(f"PARTIAL: Component 4 -- {correct_count}/{checked} cached values correct ({partial} pts)")
            total_score += partial
        elif checked == 0:
            # No cached values available (file not opened in Calc yet)
            # Fall back to checking formula structure references the correct ranges
            # BUT only award points if the formula is efficient (not nested IF)
            ws_form = wb['Performance Data']
            e2_val = ws_form['E2'].value
            if isinstance(e2_val, str) and e2_val.startswith('=') and is_efficient_formula(e2_val) and not is_nested_if(e2_val):
                # Check that formula references column D (score column)
                if 'D2' in e2_val.upper() or 'D:' in e2_val.upper() or '$D' in e2_val.upper():
                    print(f"PASS: Component 4 -- Efficient formula references score column D correctly (fallback check) (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 4 -- Formula does not reference score column D: {e2_val[:80]}")
            else:
                print(f"FAIL: Component 4 -- No efficient formula in E2 or no cached values")
        else:
            print(f"FAIL: Component 4 -- Only {correct_count}/{checked} cached values match expected categories")

        wb_data.close()
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    wb.close()

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
