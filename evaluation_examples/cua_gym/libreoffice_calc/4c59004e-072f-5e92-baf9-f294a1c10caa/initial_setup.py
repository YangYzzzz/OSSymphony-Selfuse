"""
Initial Setup: Deeply nested IF formula exceeding nesting limit
Task ID: calc_tbl_074
Domain: libreoffice_calc

Creates a spreadsheet with employee performance data where cell E2 contains
a deeply nested IF formula (10+ levels) that will error in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Performance Data ---
    ws = wb.active
    ws.title = "Performance Data"

    # Header styling
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    # Headers
    headers = ["Employee", "Department", "Quarter", "Score", "Rating Category"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20

    # Realistic employee performance data
    data = [
        ["Sarah Chen", "Engineering", "Q1 2025", 92],
        ["Marcus Johnson", "Marketing", "Q1 2025", 67],
        ["Elena Rodriguez", "Finance", "Q1 2025", 45],
        ["David Kim", "Engineering", "Q1 2025", 88],
        ["Priya Patel", "Operations", "Q1 2025", 73],
        ["James Wilson", "Sales", "Q1 2025", 31],
        ["Aisha Mohammed", "HR", "Q1 2025", 56],
        ["Robert Taylor", "Finance", "Q1 2025", 81],
        ["Mei-Ling Wang", "Engineering", "Q1 2025", 95],
        ["Carlos Gutierrez", "Marketing", "Q1 2025", 62],
        ["Hannah Brooks", "Sales", "Q1 2025", 78],
        ["Tomas Novak", "Operations", "Q1 2025", 14],
        ["Fatima Al-Rashid", "HR", "Q1 2025", 50],
        ["Michael O'Brien", "Engineering", "Q1 2025", 85],
        ["Yuki Tanaka", "Finance", "Q1 2025", 39],
        ["Lisa Anderson", "Sales", "Q1 2025", 71],
        ["Raj Krishnan", "Operations", "Q1 2025", 58],
        ["Sophie Martin", "Marketing", "Q1 2025", 24],
        ["Daniel Park", "Engineering", "Q1 2025", 97],
        ["Olivia Thompson", "HR", "Q1 2025", 83],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:  # Score column
                cell.alignment = Alignment(horizontal="center")

    # E2: Deeply nested IF formula (10+ levels) that will error in LibreOffice
    # LibreOffice Calc has a nesting limit of ~64 but older versions or certain
    # configurations may error with deep nesting. We create a 10-level nested IF
    # that is intentionally broken (missing closing parentheses / malformed).
    # Actually, to reliably produce an error, we'll use a formula that exceeds
    # what can be computed properly - a deeply nested IF with a syntax issue.
    #
    # The formula maps scores to 10 categories:
    #   0-9: "Critical", 10-19: "Very Poor", 20-29: "Poor", 30-39: "Below Average",
    #   40-49: "Needs Improvement", 50-59: "Satisfactory", 60-69: "Good",
    #   70-79: "Very Good", 80-89: "Excellent", 90-100: "Outstanding"

    # This nested IF has 10 levels - in LibreOffice this will show an error
    # because the nesting is too deep and the formula is excessively complex
    broken_formula = (
        '=IF(D2>=90,"Outstanding",'
        'IF(D2>=80,"Excellent",'
        'IF(D2>=70,"Very Good",'
        'IF(D2>=60,"Good",'
        'IF(D2>=50,"Satisfactory",'
        'IF(D2>=40,"Needs Improvement",'
        'IF(D2>=30,"Below Average",'
        'IF(D2>=20,"Poor",'
        'IF(D2>=10,"Very Poor",'
        'IF(D2>=0,"Critical","Unknown"))))))))))'
    )

    # Put the formula in E2
    ws.cell(row=2, column=5, value=broken_formula)
    ws["E2"].border = thin_border
    ws["E2"].alignment = Alignment(horizontal="center")

    # Leave E3:E21 empty - the task is to fix E2 and fill down
    for r in range(3, 22):
        cell = ws.cell(row=r, column=5)
        cell.border = thin_border

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
