"""
Initial Setup: Account number masking task
Task ID: calc_fma_replace_020
Domain: libreoffice_calc

Creates a spreadsheet with account numbers in column A and an empty
'Masked Account' column B. The agent must fill B2:B11 with REPLACE formulas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_replace_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Accounts ---
    ws = wb.active
    ws.title = 'Accounts'

    # Headers
    ws['A1'] = 'Account Number'
    ws['B1'] = 'Masked Account'

    # Style headers
    header_font = Font(name='Calibri', bold=True, size=12)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col_letter in ['A', 'B']:
        cell = ws[f'{col_letter}1']
        cell.font = Font(name='Calibri', bold=True, size=12, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align

    # Account number data (rows 2-11)
    account_numbers = [
        'ACCT-12345678',
        'ACCT-98765432',
        'ACCT-11223344',
        'ACCT-55667788',
        'ACCT-99001122',
        'ACCT-33445566',
        'ACCT-77889900',
        'ACCT-22334455',
        'ACCT-66778899',
        'ACCT-00112233',
    ]

    for i, acct in enumerate(account_numbers, start=2):
        ws.cell(row=i, column=1, value=acct)
        # Column B (Masked Account) is intentionally left empty

    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
