"""
Reward Script: Replace middle 4 characters of account numbers with asterisks using REPLACE formula
Task ID: calc_fma_replace_020
Domain: libreoffice_calc
Scoring:
  - Component 1: At least one REPLACE formula exists in B2:B11 (0.3 pts)
  - Component 2: All 10 cells B2:B11 contain correct REPLACE formulas (0.5 pts)
  - Component 3: Column A account number data is intact (0.2 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_replace_020'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    The task: Replace middle 4 characters (positions 6-9) of account numbers in column A
    with asterisks, creating masked versions using REPLACE formulas in column B (B2:B11).
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Accounts' sheet must exist
    if 'Accounts' not in wb.sheetnames:
        print("CRITICAL: 'Accounts' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Accounts']

    # Expected REPLACE formulas for B2:B11
    # Task requires =REPLACE(AX, 6, 4, "****") for rows 2-11
    expected_rows = list(range(2, 12))  # rows 2 through 11

    # Component 1: At least one REPLACE formula present in B2:B11 (0.3 points)
    # This FAILS on initial (all None) and PASSES on golden (all have formulas)
    try:
        replace_formula_count = 0
        for row in expected_rows:
            val = ws.cell(row=row, column=2).value
            if val is not None and isinstance(val, str) and 'REPLACE' in val.upper():
                replace_formula_count += 1

        if replace_formula_count >= 1:
            print(f"PASS: Component 1 — Found {replace_formula_count} REPLACE formula(s) in column B (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — No REPLACE formulas found in B2:B11 (found {replace_formula_count})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 10 cells B2:B11 contain correct REPLACE formulas (0.5 points)
    # Correct formula: =REPLACE(AX, 6, 4, "****") where X is the row number
    # This FAILS on initial (all None) and PASSES on golden (all 10 formulas present and correct)
    try:
        correct_formula_count = 0
        incorrect_details = []
        for row in expected_rows:
            val = ws.cell(row=row, column=2).value
            if val is None:
                incorrect_details.append(f"B{row}: None (empty)")
                continue
            if not isinstance(val, str):
                incorrect_details.append(f"B{row}: not a formula string (got {repr(val)})")
                continue

            formula_upper = val.upper().replace(' ', '')
            # Check for REPLACE function with correct parameters: REPLACE(AX, 6, 4, "****")
            # The formula should reference column A of the same row, start at position 6, replace 4 chars with ****
            expected_cell_ref = f'A{row}'
            has_replace = 'REPLACE(' in formula_upper
            has_correct_ref = expected_cell_ref.upper() in formula_upper
            has_pos_6 = ',6,' in formula_upper
            has_len_4 = ',4,' in formula_upper
            has_asterisks = '"****"' in val or "'****'" in val or '****' in val

            if has_replace and has_correct_ref and has_pos_6 and has_len_4 and has_asterisks:
                correct_formula_count += 1
            else:
                issues = []
                if not has_replace:
                    issues.append("missing REPLACE function")
                if not has_correct_ref:
                    issues.append(f"missing reference to {expected_cell_ref}")
                if not has_pos_6:
                    issues.append("missing start position 6")
                if not has_len_4:
                    issues.append("missing length 4")
                if not has_asterisks:
                    issues.append("missing '****' replacement string")
                incorrect_details.append(f"B{row}: {val!r} — {', '.join(issues)}")

        if correct_formula_count == 10:
            print(f"PASS: Component 2 — All 10 cells B2:B11 contain correct REPLACE formulas (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — Only {correct_formula_count}/10 cells have correct REPLACE formulas")
            for detail in incorrect_details[:5]:  # Show up to 5 failures
                print(f"  - {detail}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column A account number data is intact (0.2 points)
    # This is a compound check: awards points only when column B formulas are correct
    # (already verified in component 2) AND column A data is unmodified.
    # Because it depends on correct_formula_count == 10 (which fails on initial),
    # this compound condition also fails on initial — satisfying the litmus test.
    try:
        expected_account_numbers = [
            'ACCT-12345678', 'ACCT-98765432', 'ACCT-11223344', 'ACCT-55667788',
            'ACCT-99001122', 'ACCT-33445566', 'ACCT-77889900', 'ACCT-22334455',
            'ACCT-66778899', 'ACCT-00112233'
        ]
        col_a_mismatch_count = 0
        col_a_issues = []
        for i, row in enumerate(expected_rows):
            val = ws.cell(row=row, column=1).value
            expected = expected_account_numbers[i]
            if val != expected:
                col_a_mismatch_count += 1
                col_a_issues.append(f"A{row}: expected {expected!r}, found {repr(val)}")

        # Count any extra content in columns C+ (cells that should remain empty)
        extra_cells_count = sum(
            1
            for row in range(2, 12)
            for col in range(3, ws.max_column + 1)
            if ws.cell(row=row, column=col).value is not None
        )

        # Compound check: all 10 B-column formulas correct AND column A unmodified AND no extra columns
        if correct_formula_count == 10 and col_a_mismatch_count == 0 and extra_cells_count == 0:
            print(f"PASS: Component 3 — Column A account numbers intact and no extra modifications (0.2 pts)")
            total_score += 0.2
        elif correct_formula_count < 10:
            print(f"FAIL: Component 3 — Skipped (column B formulas incomplete, needed for compound check)")
        elif col_a_mismatch_count > 0:
            print(f"FAIL: Component 3 — Column A data was modified ({col_a_mismatch_count} mismatches):")
            for issue in col_a_issues[:3]:
                print(f"  - {issue}")
        elif extra_cells_count > 0:
            print(f"FAIL: Component 3 — {extra_cells_count} extra cells modified beyond column B")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
