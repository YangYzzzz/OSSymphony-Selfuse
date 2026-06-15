"""
Initial Setup: Monthly expense report template with partial structure
Task ID: calc_gsd_024
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Monthly Expenses'

    # Row 1: Merged title
    ws.merge_cells('A1:N1')
    ws['A1'] = 'Monthly Expense Report'
    ws['A1'].font = Font(size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Row 2: Headers
    headers = [
        'Category', 'Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
        'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec', 'Annual Total'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=2, column=col, value=h)

    # Rows 3-17: 15 expense categories in column A only (NO data in B-N)
    categories = [
        'Rent', 'Utilities', 'Insurance', 'Salaries', 'Benefits',
        'Travel', 'Meals', 'Marketing', 'Software', 'Hardware',
        'Legal', 'Consulting', 'Supplies', 'Maintenance', 'Miscellaneous'
    ]
    for r, cat in enumerate(categories, 3):
        ws.cell(row=r, column=1, value=cat)

    # Row 18 is left empty (no Grand Total yet)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col_letter].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
