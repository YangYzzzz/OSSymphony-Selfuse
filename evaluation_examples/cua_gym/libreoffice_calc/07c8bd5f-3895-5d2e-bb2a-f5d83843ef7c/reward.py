"""
Reward Script: Monthly Expense Report Template
Task ID: calc_gsd_024
Domain: libreoffice_calc
Scoring:
  1. SUM formulas in N3:N17 (row totals)           — 0.20
  2. Grand Total row (A18 label + B18:N18 formulas) — 0.25
  3. USD currency formatting on B3:N18              — 0.15
  4. Alternating gray fill on odd rows 3,5,...,17   — 0.15
  5. Borders on A2:N18                              — 0.10
  6. Bold on header row 2 and grand total row 18    — 0.10
  7. Double bottom border on row 18                 — 0.05
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_024'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Monthly Expenses' sheet must exist
    if 'Monthly Expenses' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Monthly Expenses' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Monthly Expenses']

    # Component 1: SUM formulas in N3:N17 (0.20 points)
    # Each of the 15 category rows should have =SUM(B<r>:M<r>) in column N
    try:
        sum_formula_count = 0
        for r in range(3, 18):
            val = ws.cell(row=r, column=14).value  # column N = 14
            if val and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                expected = f"=SUM(B{r}:M{r})"
                if normalized == expected:
                    sum_formula_count += 1
        if sum_formula_count == 15:
            print(f"PASS: Component 1 — All 15 row SUM formulas in N3:N17 (0.20 pts)")
            total_score += 0.20
        elif sum_formula_count > 0:
            partial = round(0.20 * sum_formula_count / 15, 4)
            print(f"PARTIAL: Component 1 — {sum_formula_count}/15 row SUM formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No SUM formulas found in N3:N17")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Grand Total row (A18 + B18:N18 formulas) (0.25 points)
    try:
        comp2_score = 0.0

        # Sub-check 2a: A18 = 'Grand Total' (0.05 points)
        a18_val = ws.cell(row=18, column=1).value
        if a18_val and str(a18_val).strip().lower() == 'grand total':
            comp2_score += 0.05
            print(f"  PASS: 2a — A18 = '{a18_val}'")
        else:
            print(f"  FAIL: 2a — A18 expected 'Grand Total', found '{a18_val}'")

        # Sub-check 2b: Column SUM formulas B18:M18 (0.10 points)
        col_sum_count = 0
        for c in range(2, 14):  # B=2 to M=13
            val = ws.cell(row=18, column=c).value
            if val and isinstance(val, str):
                col_letter = chr(ord('A') + c - 1)
                expected = f"=SUM({col_letter}3:{col_letter}17)"
                if val.upper().replace(" ", "") == expected:
                    col_sum_count += 1
        if col_sum_count == 12:
            comp2_score += 0.10
            print(f"  PASS: 2b — All 12 column SUM formulas in B18:M18")
        elif col_sum_count > 0:
            partial = round(0.10 * col_sum_count / 12, 4)
            comp2_score += partial
            print(f"  PARTIAL: 2b — {col_sum_count}/12 column SUM formulas")
        else:
            print(f"  FAIL: 2b — No column SUM formulas in B18:M18")

        # Sub-check 2c: N18 = SUM(N3:N17) (0.10 points)
        n18_val = ws.cell(row=18, column=14).value
        if n18_val and isinstance(n18_val, str):
            if n18_val.upper().replace(" ", "") == "=SUM(N3:N17)":
                comp2_score += 0.10
                print(f"  PASS: 2c — N18 = '=SUM(N3:N17)'")
            else:
                print(f"  FAIL: 2c — N18 expected '=SUM(N3:N17)', found '{n18_val}'")
        else:
            print(f"  FAIL: 2c — N18 has no formula, found '{n18_val}'")

        if comp2_score > 0:
            print(f"PASS: Component 2 — Grand Total row ({comp2_score} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — Grand Total row not set up")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: USD currency formatting on B3:N18 (0.15 points)
    # Check a sample of cells — at least 80% must have currency format
    try:
        currency_count = 0
        total_checked = 0
        for r in range(3, 19):  # rows 3-18
            for c in range(2, 15):  # cols B-N
                total_checked += 1
                nf = ws.cell(row=r, column=c).number_format
                # Accept common USD formats
                if nf and ('$' in nf or 'USD' in nf.upper()):
                    currency_count += 1

        ratio = currency_count / total_checked if total_checked > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 3 — USD currency formatting ({currency_count}/{total_checked} cells) (0.15 pts)")
            total_score += 0.15
        elif ratio >= 0.4:
            partial = round(0.15 * ratio, 4)
            print(f"PARTIAL: Component 3 — {currency_count}/{total_checked} cells with currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {currency_count}/{total_checked} cells have currency formatting")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Alternating gray fill (#F2F2F2) on odd rows 3,5,7,...,17 (0.15 points)
    # Odd rows should have FFF2F2F2 fill; even rows should NOT
    try:
        odd_rows_correct = 0
        odd_rows_total = 8  # rows 3,5,7,9,11,13,15,17
        for r in [3, 5, 7, 9, 11, 13, 15, 17]:
            try:
                fill_rgb = ws.cell(row=r, column=1).fill.fgColor.rgb
                # Accept FFF2F2F2 or 00F2F2F2 or F2F2F2 (various alpha prefixes)
                if fill_rgb and 'F2F2F2' in fill_rgb.upper():
                    odd_rows_correct += 1
            except Exception:
                pass

        # Also verify even rows do NOT have gray fill (at least some)
        even_no_gray = 0
        for r in [4, 6, 8, 10]:
            try:
                fill_rgb = ws.cell(row=r, column=1).fill.fgColor.rgb
                if fill_rgb is None or 'F2F2F2' not in fill_rgb.upper():
                    even_no_gray += 1
                elif fill_rgb == '00000000':
                    even_no_gray += 1
            except Exception:
                even_no_gray += 1  # No fill = correct

        if odd_rows_correct >= 6 and even_no_gray >= 2:
            print(f"PASS: Component 4 — Alternating gray fill ({odd_rows_correct}/{odd_rows_total} odd rows gray, {even_no_gray}/4 even rows no gray) (0.15 pts)")
            total_score += 0.15
        elif odd_rows_correct >= 3:
            partial = round(0.15 * odd_rows_correct / odd_rows_total, 4)
            print(f"PARTIAL: Component 4 — {odd_rows_correct}/{odd_rows_total} odd rows with gray fill ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {odd_rows_correct}/{odd_rows_total} odd rows have gray fill")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Borders on A2:N18 (0.10 points)
    # Sample border checks on several cells in the range
    try:
        bordered_count = 0
        sample_cells = [(2, 1), (2, 14), (3, 1), (10, 7), (17, 14), (18, 1), (18, 14)]
        for r, c in sample_cells:
            cell = ws.cell(row=r, column=c)
            has_border = False
            if cell.border:
                sides = [cell.border.left, cell.border.right, cell.border.top, cell.border.bottom]
                if any(s and s.style and s.style != 'none' for s in sides):
                    has_border = True
            if has_border:
                bordered_count += 1

        if bordered_count >= 5:
            print(f"PASS: Component 5 — Borders applied ({bordered_count}/{len(sample_cells)} sampled cells have borders) (0.10 pts)")
            total_score += 0.10
        elif bordered_count >= 2:
            partial = round(0.10 * bordered_count / len(sample_cells), 4)
            print(f"PARTIAL: Component 5 — {bordered_count}/{len(sample_cells)} sampled cells have borders ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Only {bordered_count}/{len(sample_cells)} sampled cells have borders")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Bold on row 2 (header) and row 18 (grand total) (0.10 points)
    try:
        comp6_score = 0.0

        # Check row 2 bold (0.05 pts)
        row2_bold = 0
        for c in range(1, 15):
            if ws.cell(row=2, column=c).font.bold:
                row2_bold += 1
        if row2_bold >= 10:
            comp6_score += 0.05
            print(f"  PASS: 6a — Row 2 bold ({row2_bold}/14 cells)")
        else:
            print(f"  FAIL: 6a — Row 2 bold only {row2_bold}/14 cells")

        # Check row 18 bold (0.05 pts)
        row18_bold = 0
        for c in range(1, 15):
            if ws.cell(row=18, column=c).font.bold:
                row18_bold += 1
        if row18_bold >= 10:
            comp6_score += 0.05
            print(f"  PASS: 6b — Row 18 bold ({row18_bold}/14 cells)")
        else:
            print(f"  FAIL: 6b — Row 18 bold only {row18_bold}/14 cells")

        if comp6_score > 0:
            print(f"PASS: Component 6 — Bold formatting ({comp6_score} pts)")
            total_score += comp6_score
        else:
            print(f"FAIL: Component 6 — No bold formatting on rows 2 or 18")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Double bottom border on row 18 (0.05 points)
    try:
        double_border_count = 0
        for c in range(1, 15):
            cell = ws.cell(row=18, column=c)
            if cell.border and cell.border.bottom and cell.border.bottom.style == 'double':
                double_border_count += 1

        if double_border_count >= 10:
            print(f"PASS: Component 7 — Double bottom border on row 18 ({double_border_count}/14 cells) (0.05 pts)")
            total_score += 0.05
        elif double_border_count >= 1:
            # At minimum N18 has it per task spec
            partial = round(0.05 * min(double_border_count / 10, 1.0), 4)
            print(f"PARTIAL: Component 7 — {double_border_count}/14 cells with double bottom border ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 7 — No double bottom border on row 18")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification (LibreOffice may have unsaved edits)
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    persist_app_state()
    verify_task(file_path)
