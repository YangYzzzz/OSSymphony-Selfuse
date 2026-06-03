"""
Reward Script: GL Account Summary with SUMIF formulas, named ranges, conditional formatting
Task ID: calc_fin_gl_account_summary_074
Domain: libreoffice_calc
Scoring:
  - Component 1: SUMIF formulas in C2:C15 and D2:D15 (0.30 pts)
  - Component 2: Net balance formulas E2:E15 (0.15 pts)
  - Component 3: Row 16 SUM totals with bold formatting and currency format (0.15 pts)
  - Component 4: Named range GL_Data on GL_Transactions covering $A$2:$E$200 (0.15 pts)
  - Component 5: Conditional formatting on E2:E15 ABS(E2)>100000 with yellow background (0.15 pts)
  - Component 6: Header row 1 bold + GL_Summary freeze_panes=A2 (0.10 pts)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_gl_account_summary_074'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: required sheets must exist
    if 'GL_Summary' not in wb.sheetnames or 'GL_Transactions' not in wb.sheetnames:
        print(f"CRITICAL: Required sheets not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_summary = wb['GL_Summary']

    # Component 1: SUMIF formulas in C2:C15 and D2:D15 (0.30 points)
    # C column: =SUMIF(GL_Transactions.$B$2:$B$200,Ax,GL_Transactions.$D$2:$D$200)
    # D column: =SUMIF(GL_Transactions.$B$2:$B$200,Ax,GL_Transactions.$E$2:$E$200)
    try:
        c_sumif_count = 0
        d_sumif_count = 0
        expected_rows = 14  # rows 2 through 15

        for row in range(2, 16):
            cell_c = ws_summary.cell(row=row, column=3)
            cell_d = ws_summary.cell(row=row, column=4)

            # Check C column SUMIF formula
            if cell_c.value and isinstance(cell_c.value, str):
                val_upper = cell_c.value.upper().replace(' ', '')
                if 'SUMIF' in val_upper and 'GL_TRANSACTIONS' in val_upper and '$D$2:$D$200' in val_upper:
                    c_sumif_count += 1

            # Check D column SUMIF formula
            if cell_d.value and isinstance(cell_d.value, str):
                val_upper = cell_d.value.upper().replace(' ', '')
                if 'SUMIF' in val_upper and 'GL_TRANSACTIONS' in val_upper and '$E$2:$E$200' in val_upper:
                    d_sumif_count += 1

        if c_sumif_count == expected_rows and d_sumif_count == expected_rows:
            print(f"PASS: Component 1 — All {expected_rows} SUMIF formulas in C2:C15 and D2:D15 (0.30 pts)")
            total_score += 0.30
        elif c_sumif_count > 0 or d_sumif_count > 0:
            # Partial credit for partial formula coverage
            partial = (c_sumif_count + d_sumif_count) / (expected_rows * 2) * 0.30
            print(f"PARTIAL: Component 1 — C SUMIF: {c_sumif_count}/{expected_rows}, D SUMIF: {d_sumif_count}/{expected_rows} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No SUMIF formulas found in C2:C15 or D2:D15")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Net balance formulas E2:E15 (=Cx-Dx) (0.15 points)
    try:
        e_formula_count = 0
        expected_rows = 14  # rows 2 through 15

        for row in range(2, 16):
            cell_e = ws_summary.cell(row=row, column=5)
            if cell_e.value and isinstance(cell_e.value, str):
                val_upper = cell_e.value.upper().replace(' ', '')
                # Formula should be =C{row}-D{row}
                expected = f'=C{row}-D{row}'.upper().replace(' ', '')
                if val_upper == expected:
                    e_formula_count += 1

        if e_formula_count == expected_rows:
            print(f"PASS: Component 2 — All {expected_rows} net balance formulas E2:E15 (=Cx-Dx) (0.15 pts)")
            total_score += 0.15
        elif e_formula_count > 0:
            partial = (e_formula_count / expected_rows) * 0.15
            print(f"PARTIAL: Component 2 — {e_formula_count}/{expected_rows} net balance formulas in E2:E15 ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No net balance formulas (=Cx-Dx) found in E2:E15")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row 16 SUM totals with bold formatting and currency format (0.15 points)
    # C16=SUM(C2:C15), D16=SUM(D2:D15), E16=SUM(E2:E15), all bold
    try:
        c16 = ws_summary.cell(row=16, column=3)
        d16 = ws_summary.cell(row=16, column=4)
        e16 = ws_summary.cell(row=16, column=5)

        c16_ok = (c16.value and isinstance(c16.value, str) and
                  'SUM(C2:C15)' in c16.value.upper().replace(' ', ''))
        d16_ok = (d16.value and isinstance(d16.value, str) and
                  'SUM(D2:D15)' in d16.value.upper().replace(' ', ''))
        e16_ok = (e16.value and isinstance(e16.value, str) and
                  'SUM(E2:E15)' in e16.value.upper().replace(' ', ''))
        bold_ok = (c16.font.bold == True and d16.font.bold == True and e16.font.bold == True)

        if c16_ok and d16_ok and e16_ok and bold_ok:
            print(f"PASS: Component 3 — Row 16 SUM totals present and bold (0.15 pts)")
            total_score += 0.15
        elif c16_ok and d16_ok and e16_ok:
            print(f"PARTIAL: Component 3 — Row 16 SUM totals present but not all bold (0.08 pts)")
            total_score += 0.08
        elif c16_ok or d16_ok or e16_ok:
            print(f"PARTIAL: Component 3 — Only some row 16 totals present (0.05 pts): C16={c16_ok}, D16={d16_ok}, E16={e16_ok}")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No row 16 SUM total formulas found. C16={repr(c16.value)}, D16={repr(d16.value)}, E16={repr(e16.value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Named range GL_Data on GL_Transactions covering $A$2:$E$200 (0.15 points)
    try:
        found_gl_data = False
        if 'GL_Data' in wb.defined_names:
            defn = wb.defined_names['GL_Data']
            try:
                dests = list(defn.destinations)
                for sheet_title, coord in dests:
                    if sheet_title == 'GL_Transactions' and '$A$2:$E$200' in coord.upper():
                        found_gl_data = True
                        break
            except Exception as e2:
                print(f"  Named range destinations error: {e2}")

        if found_gl_data:
            print(f"PASS: Component 4 — Named range 'GL_Data' found on GL_Transactions sheet covering $A$2:$E$200 (0.15 pts)")
            total_score += 0.15
        elif 'GL_Data' in wb.defined_names:
            print(f"PARTIAL: Component 4 — Named range 'GL_Data' found but destinations don't match expected. (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — Named range 'GL_Data' not found. Available: {list(wb.defined_names)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on E2:E15 with ABS(E2)>100000 and yellow background (0.15 points)
    try:
        cf_found = False
        cf_formula_ok = False
        cf_yellow_ok = False

        cf_rules = ws_summary.conditional_formatting
        for cf in cf_rules:
            cf_range_str = str(cf).upper()
            # Check if range includes E2:E15
            if 'E2' in cf_range_str and 'E15' in cf_range_str:
                for rule in cf.rules:
                    cf_found = True
                    # Check formula
                    if hasattr(rule, 'formula') and rule.formula:
                        for formula in rule.formula:
                            f_upper = str(formula).upper().replace(' ', '')
                            if 'ABS(E2)' in f_upper and '100000' in f_upper:
                                cf_formula_ok = True
                    # Check yellow fill
                    if hasattr(rule, 'dxf') and rule.dxf:
                        dxf = rule.dxf
                        if hasattr(dxf, 'fill') and dxf.fill:
                            try:
                                fg_rgb = dxf.fill.fgColor.rgb
                                # Yellow: FFFFFF00 or FFFF00 variants
                                if fg_rgb in ('FFFFFF00', '00FFFF00') or 'FFFF00' in fg_rgb.upper():
                                    cf_yellow_ok = True
                            except Exception:
                                pass

        if cf_found and cf_formula_ok and cf_yellow_ok:
            print(f"PASS: Component 5 — Conditional formatting on E2:E15 with ABS formula and yellow background (0.15 pts)")
            total_score += 0.15
        elif cf_found and cf_formula_ok:
            print(f"PARTIAL: Component 5 — CF found with correct formula but no yellow background (0.08 pts)")
            total_score += 0.08
        elif cf_found:
            print(f"PARTIAL: Component 5 — CF found on E range but formula or color mismatch (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting found on E2:E15")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Header row 1 bold + GL_Summary freeze_panes=A2 (0.10 points)
    try:
        row1_bold = all(
            ws_summary.cell(row=1, column=col).font.bold == True
            for col in range(1, 6)
        )
        freeze_ok = (ws_summary.freeze_panes == 'A2')

        if row1_bold and freeze_ok:
            print(f"PASS: Component 6 — Row 1 headers bold and freeze_panes=A2 (0.10 pts)")
            total_score += 0.10
        elif row1_bold:
            print(f"PARTIAL: Component 6 — Row 1 bold but freeze_panes not set (found: {ws_summary.freeze_panes}) (0.05 pts)")
            total_score += 0.05
        elif freeze_ok:
            print(f"PARTIAL: Component 6 — freeze_panes=A2 but row 1 not bold (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Row 1 bold={row1_bold}, freeze_panes={ws_summary.freeze_panes}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
