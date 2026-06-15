"""
Initial Setup: GL Account Summary with SUMIF formulas and conditional formatting
Task ID: calc_fin_gl_account_summary_074
Domain: libreoffice_calc

Creates:
- GL_Transactions sheet with 199 realistic journal entries (rows 2-200)
- GL_Summary sheet with 14 GL accounts (A2:B15 filled), C/D/E empty
- No named ranges, no formulas in GL_Summary, no formatting
"""

import openpyxl
from openpyxl.styles import Font
import random
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_gl_account_summary_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# GL accounts: number -> name
GL_ACCOUNTS = [
    ('1000', 'Cash and Cash Equivalents'),
    ('1100', 'Accounts Receivable'),
    ('1200', 'Inventory'),
    ('1300', 'Prepaid Expenses'),
    ('1500', 'Property and Equipment'),
    ('2000', 'Accounts Payable'),
    ('2100', 'Accrued Liabilities'),
    ('2200', 'Short-Term Debt'),
    ('3000', 'Common Stock'),
    ('3100', 'Retained Earnings'),
    ('4000', 'Revenue'),
    ('5000', 'Cost of Goods Sold'),
    ('6000', 'Selling Expenses'),
    ('6100', 'General and Administrative'),
]

DESCRIPTIONS = [
    'Customer payment received',
    'Vendor invoice payment',
    'Payroll disbursement',
    'Office supplies purchase',
    'Sales revenue recognition',
    'Inventory purchase',
    'Equipment maintenance',
    'Rent payment',
    'Utility bill payment',
    'Professional services fee',
    'Marketing expense',
    'Travel reimbursement',
    'Insurance premium',
    'Software subscription',
    'Bank charges',
    'Interest income',
    'Depreciation charge',
    'Tax payment',
    'Consulting fee',
    'Customer refund issued',
    'Loan repayment',
    'Dividend payment',
    'Prepaid expense amortization',
    'Accrual adjustment',
    'Freight and shipping',
]

def create_initial():
    random.seed(42)  # reproducible data

    wb = openpyxl.Workbook()

    # --- Sheet 1: GL_Transactions ---
    ws_tx = wb.active
    ws_tx.title = 'GL_Transactions'

    # Headers (row 1) - NOT bold, no freeze in initial
    tx_headers = ['Date', 'GL Account', 'Description', 'Debit', 'Credit']
    for col, h in enumerate(tx_headers, 1):
        ws_tx.cell(row=1, column=col, value=h)

    # Generate 199 realistic journal entries (rows 2-200)
    start_date = date(2024, 1, 1)
    account_numbers = [acct[0] for acct in GL_ACCOUNTS]

    for row in range(2, 201):
        # Random date within 2024
        offset = random.randint(0, 364)
        entry_date = start_date + timedelta(days=offset)

        # Pick a GL account
        acct_num = random.choice(account_numbers)

        # Pick a description
        desc = random.choice(DESCRIPTIONS)

        # Generate debit/credit amounts (realistic business values)
        # Most entries are either debit or credit, not both
        amount = round(random.uniform(250, 85000), 2)
        if random.random() < 0.5:
            debit = amount
            credit = 0.0
        else:
            debit = 0.0
            credit = amount

        ws_tx.cell(row=row, column=1, value=entry_date.strftime('%Y-%m-%d'))
        ws_tx.cell(row=row, column=2, value=acct_num)
        ws_tx.cell(row=row, column=3, value=desc)
        ws_tx.cell(row=row, column=4, value=debit if debit > 0 else None)
        ws_tx.cell(row=row, column=5, value=credit if credit > 0 else None)

    # --- Sheet 2: GL_Summary ---
    ws_sum = wb.create_sheet('GL_Summary')

    # Headers (row 1) - NOT bold, no freeze in initial
    sum_headers = ['GL Account', 'Account Name', 'Total Debits', 'Total Credits', 'Net Balance']
    for col, h in enumerate(sum_headers, 1):
        ws_sum.cell(row=1, column=col, value=h)

    # A2:B15: 14 GL accounts already entered
    for i, (acct_num, acct_name) in enumerate(GL_ACCOUNTS, 2):
        ws_sum.cell(row=i, column=1, value=acct_num)
        ws_sum.cell(row=i, column=2, value=acct_name)
    # C, D, E columns are intentionally empty (task: add SUMIF formulas)

    # Adjust column widths for readability
    ws_tx.column_dimensions['A'].width = 12
    ws_tx.column_dimensions['B'].width = 12
    ws_tx.column_dimensions['C'].width = 35
    ws_tx.column_dimensions['D'].width = 14
    ws_tx.column_dimensions['E'].width = 14

    ws_sum.column_dimensions['A'].width = 14
    ws_sum.column_dimensions['B'].width = 30
    ws_sum.column_dimensions['C'].width = 15
    ws_sum.column_dimensions['D'].width = 15
    ws_sum.column_dimensions['E'].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: GL_Transactions (199 rows), GL_Summary (14 GL accounts, C/D/E empty)')

create_initial()
