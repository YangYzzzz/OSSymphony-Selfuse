"""
FINAL REWARD SCRIPT - SUCCESS
Task: I need a new column for total cost that references the 'Cost Sheet' for per-unit costs (multiply by quantity and subtract discounts). Then generate a Pivot Table in a new sheet summarizing costs by supplier.
Generated: 2025-11-24 07:34:01
Status: success
Model: o3
Total Steps: 8
"""

import openpyxl
import math

############################################################
# Reward Script for:                                       #
# 1) Creating “Total Cost” column in ‘Orders’ sheet that   #
#    references ‘Cost Sheet’ and calculates                #
#    Qty*Cost-per-Unit – Discount                          #
# 2) Producing a new sheet (pivot / summary) that shows     #
#    total costs grouped by Supplier                       #
############################################################

FILE_PATH = "/home/user/i_need_a_new_column_for_total_cost_that_references_the_cost_sheet_for_per_unit_costs_multiply_by_qua.xlsx"

# -----------------  CORE VERIFICATION HELPERS -----------------

def safe_float(val):
    """Attempt to convert a value to float, return None if impossible."""
    try:
        return float(val)
    except Exception:
        return None

# -----------------  VERIFICATION 1: TOTAL COST COLUMN --------

def verify_total_cost_column(wb):
    """Verify presence & correctness of Total Cost column and formulas."""
    score = 0.0            # up to 0.5 points
    details = []

    # ---- Presence of required sheets ----
    if "Orders" not in wb.sheetnames or "Cost Sheet" not in wb.sheetnames:
        details.append("Missing mandatory sheets ‘Orders’ or ‘Cost Sheet’.")
        return score, details

    orders_ws = wb["Orders"]
    cost_ws   = wb["Cost Sheet"]

    # ---- Build cost-per-unit lookup dict ----
    cost_lookup = {}
    for item_cell, cost_cell in cost_ws.iter_rows(min_row=2, max_col=2, values_only=True):
        if item_cell is None:   # skip blank rows
            continue
        cost_val = safe_float(cost_cell)
        if cost_val is not None:
            cost_lookup[str(item_cell).strip()] = cost_val

    # ---- Identify columns ----
    headers = [c.value for c in orders_ws[1]]
    header_map = {str(h): idx for idx, h in enumerate(headers)}

    required_cols = ["Item", "Quantity", "Discount", "Total Cost"]
    for col in required_cols:
        if col not in header_map:
            details.append(f"Column ‘{col}’ missing in Orders sheet.")
            return score, details

    total_cost_idx = header_map["Total Cost"] + 1  # 1-based index for openpyxl
    item_idx       = header_map["Item"]        + 1
    qty_idx        = header_map["Quantity"]    + 1
    disc_idx       = header_map["Discount"]    + 1

    # ---- Load data-only copy for numeric comparison ----
    wb_data_only = openpyxl.load_workbook(FILE_PATH, data_only=True)
    orders_vals  = wb_data_only["Orders"]

    all_rows_correct = True
    checked_rows     = 0

    for r in range(2, orders_ws.max_row + 1):
        item = orders_ws.cell(r, item_idx).value
        qty  = safe_float(orders_ws.cell(r, qty_idx).value)
        disc = safe_float(orders_ws.cell(r, disc_idx).value)

        if item is None or qty is None or disc is None:
            continue  # skip malformed rows

        cpu  = cost_lookup.get(str(item).strip())
        if cpu is None:
            all_rows_correct = False
            details.append(f"No cost-per-unit entry for item ‘{item}’.")
            continue

        expected = qty * cpu - disc
        actual   = safe_float(orders_vals.cell(r, total_cost_idx).value)

        if actual is None or not math.isclose(actual, expected, rel_tol=1e-2, abs_tol=1):
            all_rows_correct = False
            details.append(f"Row {r}: expected {expected}, found {actual}.")
        checked_rows += 1

    # ---- Scoring ----
    # 0.2 for column presence, 0.2 for correct numeric results, 0.1 for proper Cost Sheet reference
    score += 0.2  # column exists (already ensured)

    if checked_rows > 0 and all_rows_correct:
        score += 0.2
    else:
        details.append("One or more Total Cost values incorrect.")

    # ---- Formula reference check (must mention ‘Cost Sheet’) ----
    ref_ok = 0
    formula_rows = 0
    for r in range(2, orders_ws.max_row + 1):
        cell_val = orders_ws.cell(r, total_cost_idx).value
        if isinstance(cell_val, str) and cell_val.startswith("="):
            formula_rows += 1
            if "COST SHEET" in cell_val.upper():
                ref_ok += 1
    if formula_rows and ref_ok / formula_rows >= 0.5:
        score += 0.1
    else:
        details.append("Less than half of Total Cost formulas reference ‘Cost Sheet’.")

    return score, details

# -----------------  VERIFICATION 2: SUPPLIER SUMMARY SHEET ----

def verify_supplier_summary(wb):
    """Verify existence and correctness of pivot/summary by Supplier."""
    score = 0.0          # up to 0.5 points
    details = []

    # ---- Locate summary sheet (any sheet besides the originals) ----
    summary_ws = None
    for name in wb.sheetnames:
        if name not in ("Orders", "Cost Sheet"):
            summary_ws = wb[name]
            break

    if summary_ws is None:
        details.append("No supplier summary / pivot sheet found.")
        return score, details

    score += 0.2  # sheet exists

    # ---- Expected totals per supplier (re-compute) ----
    orders_ws = wb["Orders"]
    hdrs = [c.value for c in orders_ws[1]]
    idx_item = hdrs.index("Item") + 1
    idx_qty  = hdrs.index("Quantity") + 1
    idx_disc = hdrs.index("Discount") + 1
    idx_sup  = hdrs.index("Supplier") + 1
    idx_tot  = hdrs.index("Total Cost") + 1

    wb_data = openpyxl.load_workbook(FILE_PATH, data_only=True)
    orders_vals = wb_data["Orders"]

    supplier_totals = {}
    for r in range(2, orders_ws.max_row + 1):
        supplier = orders_ws.cell(r, idx_sup).value
        actual   = safe_float(orders_vals.cell(r, idx_tot).value)
        if supplier is None or actual is None:
            continue
        supplier_totals[supplier] = supplier_totals.get(supplier, 0) + actual

    # ---- Scan summary sheet for matching totals ----
    headers_summary = [c.value for c in summary_ws[1]] if summary_ws.max_row >= 1 else []
    if "Supplier" not in headers_summary:
        details.append("Summary sheet lacks ‘Supplier’ header.")
        return score, details

    supp_col = headers_summary.index("Supplier") + 1
    matched = 0

    for supp, exp_total in supplier_totals.items():
        found_row = None
        for r in range(2, summary_ws.max_row + 1):
            cell_val = summary_ws.cell(r, supp_col).value
            if isinstance(cell_val, str) and cell_val.strip().lower() == supp.lower():
                found_row = r
                break
        if found_row:
            # look for any numeric in that row
            numeric_val = None
            for c in summary_ws.iter_cols(min_row=found_row, max_row=found_row, min_col=1, max_col=summary_ws.max_column, values_only=True):
                for v in c:
                    if isinstance(v, (int, float)):
                        numeric_val = float(v)
                        break
                if numeric_val is not None:
                    break
            if numeric_val is not None and math.isclose(numeric_val, exp_total, rel_tol=1e-2, abs_tol=1):
                matched += 1
    if matched == len(supplier_totals):
        score += 0.3  # all suppliers match
    else:
        score += 0.3 * (matched / max(1, len(supplier_totals)))
        details.append(f"Summary matches {matched}/{len(supplier_totals)} suppliers.")

    return score, details

# -----------------  MAIN VERIFICATION WRAPPER ---------------

def verify_task(file_path):
    print("Starting verification of LibreOffice Calc task…")

    try:
        wb = openpyxl.load_workbook(file_path, data_only=False)
    except Exception as e:
        print(f"✗ Could not open workbook: {e}")
        print("REWARD: 0.0")
        return 0.0

    grand_score = 0.0
    all_details = []

    # --- Check Total Cost column (0.5) ---
    s1, d1 = verify_total_cost_column(wb)
    grand_score += s1
    all_details.extend(d1)

    # --- Check Supplier summary sheet (0.5) ---
    s2, d2 = verify_supplier_summary(wb)
    grand_score += s2
    all_details.extend(d2)

    # Ensure within [0,1]
    grand_score = round(min(1.0, grand_score), 4)

    # ---- Output diagnostics ----
    print("---------------- DETAIL REPORT ----------------")
    for d in all_details:
        print(" -", d)
    print("----------------------------------------------")
    print(f"Total Verification Score: {grand_score}")
    print(f"REWARD: {grand_score}")
    return grand_score

# --------------- Run when executed as script ----------------
if __name__ == "__main__":
    verify_task(FILE_PATH)

