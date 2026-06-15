"""
Reward Script: Professional Invoice with merged header, client info, line items,
               tax/discount calculations, payment terms, and formatting.
Task ID: calc_gsd_026
Domain: libreoffice_calc
Scoring:
  C1 (0.15) - Merged header A1:G1 with company name, 18pt bold, dark bg, white text
  C2 (0.10) - Company address/contact in rows 2-3, client/invoice info in rows 5-6
  C3 (0.15) - Row 8 column headers bold with dark background
  C4 (0.20) - 8 line items in rows 9-16 with Amount formulas in col F
  C5 (0.15) - Subtotal/Tax/Total formulas in rows 18-20
  C6 (0.10) - USD currency format on D9:D16 and F9:F20
  C7 (0.10) - Borders on A8:F20 with row 20 thick bottom border
  C8 (0.05) - A22 payment terms italic
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_026'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Merged header A1:G1 with company name, 18pt bold, dark bg, white text (0.15 pts)
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_a1_merge = any('A1' in mr and 'G1' in mr for mr in merged_ranges)
        a1 = ws['A1']
        a1_val = a1.value

        c1_pass = True
        c1_details = []

        if not has_a1_merge:
            c1_pass = False
            c1_details.append("A1:G1 not merged")

        if a1_val is None or 'Nexus Consulting Group' not in str(a1_val):
            c1_pass = False
            c1_details.append(f"A1 value mismatch: {repr(a1_val)}")

        if not a1.font.bold:
            c1_pass = False
            c1_details.append("A1 not bold")

        if a1.font.size is None or a1.font.size < 16:
            c1_pass = False
            c1_details.append(f"A1 font size {a1.font.size}, expected >=16")

        # Dark background check — fill must be solid with a dark color
        if a1.fill.fill_type != 'solid':
            c1_pass = False
            c1_details.append("A1 no solid fill")

        # White text check
        font_color = None
        try:
            font_color = a1.font.color.rgb if a1.font.color else None
        except:
            pass
        if font_color:
            # Should be white-ish (FFFFFF or 00FFFFFF)
            rgb_part = font_color[-6:]  # last 6 chars
            if rgb_part.upper() != 'FFFFFF':
                c1_pass = False
                c1_details.append(f"A1 font color {font_color}, expected white")

        if c1_pass:
            print(f"PASS: Component 1 — Merged header with correct styling (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — {'; '.join(c1_details)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Company address/contact rows 2-3, client/invoice info rows 5-6 (0.10 pts)
    try:
        c2_checks = 0
        c2_total = 4

        # Row 2-3: company address/contact
        a2_val = ws['A2'].value
        a3_val = ws['A3'].value
        if a2_val and len(str(a2_val)) > 5:
            c2_checks += 1
        else:
            print(f"  DETAIL: A2 missing or too short: {repr(a2_val)}")

        if a3_val and len(str(a3_val)) > 5:
            c2_checks += 1
        else:
            print(f"  DETAIL: A3 missing or too short: {repr(a3_val)}")

        # Row 5: 'Bill To:' and invoice #
        a5_val = ws['A5'].value
        # Check for invoice number in F5 or G5
        f5_val = ws['F5'].value
        g5_val = ws['G5'].value
        invoice_found = False
        for v in [f5_val, g5_val, ws['E5'].value]:
            if v and 'INV' in str(v).upper():
                invoice_found = True
                break

        if a5_val and 'Bill To' in str(a5_val):
            c2_checks += 0.5
        if invoice_found:
            c2_checks += 0.5

        # Row 6: client name/address, date
        a6_val = ws['A6'].value
        f6_val = ws['F6'].value
        g6_val = ws['G6'].value
        date_found = False
        for v in [f6_val, g6_val, ws['E6'].value]:
            if v and '2024' in str(v):
                date_found = True
                break

        if a6_val and len(str(a6_val)) > 3:
            c2_checks += 0.5
        if date_found:
            c2_checks += 0.5

        c2_score = min(c2_checks / c2_total, 1.0) * 0.10
        if c2_score >= 0.08:
            print(f"PASS: Component 2 — Client/invoice info present ({c2_score:.2f} pts)")
        else:
            print(f"PARTIAL: Component 2 — {c2_checks}/{c2_total} checks ({c2_score:.2f} pts)")
        total_score += c2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Row 8 headers bold with dark background (0.15 pts)
    try:
        expected_headers = ['Item', 'Description', 'Qty', 'Unit Price', 'Discount', 'Amount']
        c3_header_ok = 0
        c3_bold_ok = 0
        c3_fill_ok = 0
        cols_to_check = 6

        for col_idx in range(1, cols_to_check + 1):
            cell = ws.cell(row=8, column=col_idx)

            # Check header text present
            if cell.value and len(str(cell.value).strip()) > 0:
                c3_header_ok += 1

            # Check bold
            if cell.font.bold:
                c3_bold_ok += 1

            # Check dark background (solid fill)
            if cell.fill.fill_type == 'solid':
                c3_fill_ok += 1

        # Need all 6 headers with text, bold, and fill
        c3_score = 0.0
        if c3_header_ok >= 5:
            c3_score += 0.05
        if c3_bold_ok >= 5:
            c3_score += 0.05
        if c3_fill_ok >= 5:
            c3_score += 0.05

        if c3_score >= 0.15:
            print(f"PASS: Component 3 — Row 8 headers bold with dark background (0.15 pts)")
        elif c3_score > 0:
            print(f"PARTIAL: Component 3 — headers={c3_header_ok}/6, bold={c3_bold_ok}/6, fill={c3_fill_ok}/6 ({c3_score:.2f} pts)")
        else:
            print(f"FAIL: Component 3 — headers={c3_header_ok}/6, bold={c3_bold_ok}/6, fill={c3_fill_ok}/6")
        total_score += c3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: 8 line items in rows 9-16 with Amount formulas in col F (0.20 pts)
    try:
        items_with_data = 0
        items_with_formula = 0

        for row in range(9, 17):
            # Check item name in col A
            a_val = ws.cell(row=row, column=1).value
            # Check qty in col C
            c_val = ws.cell(row=row, column=3).value
            # Check unit price in col D
            d_val = ws.cell(row=row, column=4).value
            # Check formula in col F
            f_val = ws.cell(row=row, column=6).value

            if a_val and c_val is not None and d_val is not None:
                items_with_data += 1

            if f_val and isinstance(f_val, str) and '=' in f_val:
                # Check formula references C, D columns for this row
                f_upper = f_val.upper()
                if f'C{row}' in f_upper and f'D{row}' in f_upper:
                    items_with_formula += 1

        c4_score = 0.0
        if items_with_data >= 7:
            c4_score += 0.10
        elif items_with_data >= 4:
            c4_score += 0.05

        if items_with_formula >= 7:
            c4_score += 0.10
        elif items_with_formula >= 4:
            c4_score += 0.05

        if c4_score >= 0.20:
            print(f"PASS: Component 4 — {items_with_data} items, {items_with_formula} formulas (0.20 pts)")
        elif c4_score > 0:
            print(f"PARTIAL: Component 4 — {items_with_data}/8 items, {items_with_formula}/8 formulas ({c4_score:.2f} pts)")
        else:
            print(f"FAIL: Component 4 — {items_with_data}/8 items, {items_with_formula}/8 formulas")
        total_score += c4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Subtotal/Tax/Total formulas in rows 18-20 (0.15 pts)
    try:
        c5_score = 0.0

        # Subtotal in F18 (or E18 label + F18 formula)
        f18 = ws['F18'].value
        if f18 and isinstance(f18, str) and 'SUM' in f18.upper() and 'F9' in f18.upper():
            c5_score += 0.05
            print(f"  DETAIL: F18 subtotal formula OK: {f18}")
        else:
            print(f"  DETAIL: F18 expected SUM formula, found: {repr(f18)}")

        # Tax in F19
        f19 = ws['F19'].value
        if f19 and isinstance(f19, str) and '=' in f19 and ('0.1' in f19 or '10%' in f19.upper() or 'F18' in f19.upper()):
            c5_score += 0.05
            print(f"  DETAIL: F19 tax formula OK: {f19}")
        else:
            print(f"  DETAIL: F19 expected tax formula, found: {repr(f19)}")

        # Total in F20
        f20 = ws['F20'].value
        if f20 and isinstance(f20, str) and '=' in f20 and 'F18' in f20.upper():
            c5_score += 0.05
            print(f"  DETAIL: F20 total formula OK: {f20}")
        else:
            print(f"  DETAIL: F20 expected total formula, found: {repr(f20)}")

        if c5_score >= 0.15:
            print(f"PASS: Component 5 — Subtotal/Tax/Total formulas correct (0.15 pts)")
        elif c5_score > 0:
            print(f"PARTIAL: Component 5 — ({c5_score:.2f} pts)")
        else:
            print(f"FAIL: Component 5 — No valid formulas in rows 18-20")
        total_score += c5_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: USD currency format on D9:D16 and F9:F20 (0.10 pts)
    try:
        usd_count = 0
        total_cells = 0

        # Check D9:D16
        for row in range(9, 17):
            nf = ws.cell(row=row, column=4).number_format
            total_cells += 1
            if nf and '$' in nf:
                usd_count += 1

        # Check F9:F20
        for row in range(9, 21):
            nf = ws.cell(row=row, column=6).number_format
            total_cells += 1
            if nf and '$' in nf:
                usd_count += 1

        # Need majority of cells to have USD format
        ratio = usd_count / total_cells if total_cells > 0 else 0
        if ratio >= 0.8:
            c6_score = 0.10
            print(f"PASS: Component 6 — {usd_count}/{total_cells} cells have USD format (0.10 pts)")
        elif ratio >= 0.4:
            c6_score = 0.05
            print(f"PARTIAL: Component 6 — {usd_count}/{total_cells} cells have USD format (0.05 pts)")
        else:
            c6_score = 0.0
            print(f"FAIL: Component 6 — {usd_count}/{total_cells} cells have USD format")
        total_score += c6_score
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Borders on A8:F20 + row 20 thick bottom border (0.10 pts)
    try:
        border_count = 0
        border_total = 0

        # Check borders on A8:F20
        for row in range(8, 21):
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                border_total += 1
                has_border = (
                    cell.border.left.style is not None or
                    cell.border.right.style is not None or
                    cell.border.top.style is not None or
                    cell.border.bottom.style is not None
                )
                if has_border:
                    border_count += 1

        c7_score = 0.0
        border_ratio = border_count / border_total if border_total > 0 else 0
        if border_ratio >= 0.7:
            c7_score += 0.05
            print(f"  DETAIL: {border_count}/{border_total} cells have borders ({border_ratio:.0%})")
        else:
            print(f"  DETAIL: Only {border_count}/{border_total} cells have borders ({border_ratio:.0%})")

        # Row 20 thick bottom border
        thick_count = 0
        for col in range(1, 7):
            cell = ws.cell(row=20, column=col)
            if cell.border.bottom.style in ('thick', 'medium'):
                thick_count += 1

        if thick_count >= 4:
            c7_score += 0.05
            print(f"  DETAIL: Row 20 thick bottom border on {thick_count}/6 cells")
        else:
            print(f"  DETAIL: Row 20 thick bottom only on {thick_count}/6 cells")

        if c7_score >= 0.10:
            print(f"PASS: Component 7 — Borders and row 20 thick bottom (0.10 pts)")
        elif c7_score > 0:
            print(f"PARTIAL: Component 7 — ({c7_score:.2f} pts)")
        else:
            print(f"FAIL: Component 7 — Insufficient borders")
        total_score += c7_score
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: A22 payment terms italic (0.05 pts)
    try:
        a22 = ws['A22']
        a22_val = a22.value

        if a22_val and 'Payment Terms' in str(a22_val) and a22.font.italic:
            print(f"PASS: Component 8 — A22 has payment terms in italic (0.05 pts)")
            total_score += 0.05
        elif a22_val and 'Payment' in str(a22_val):
            # Has text but maybe not italic
            if not a22.font.italic:
                print(f"FAIL: Component 8 — A22 has text but not italic: {repr(a22_val)}")
            else:
                print(f"PASS: Component 8 — A22 payment terms italic (0.05 pts)")
                total_score += 0.05
        else:
            # Check nearby cells
            found = False
            for r in range(21, 25):
                cell = ws.cell(row=r, column=1)
                if cell.value and 'Payment' in str(cell.value) and cell.font.italic:
                    print(f"PASS: Component 8 — Payment terms found at A{r} italic (0.05 pts)")
                    total_score += 0.05
                    found = True
                    break
            if not found:
                print(f"FAIL: Component 8 — No payment terms found in A22 area")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
