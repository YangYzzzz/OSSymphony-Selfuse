"""
Initial Setup: Product list with AutoFilter enabled, no filters applied
Task ID: calc_dop_filter_beginswith_013
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_filter_beginswith_013'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'

    # Headers in row 1
    headers = ['Product ID', 'Product Name', 'Category', 'Price', 'Stock']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 120 rows of data (rows 2-121)
    # ~14 products start with "Pro", rest are other products
    # Order: mix them so Pro products are spread across the list

    all_products = [
        # Rows 2-11 (non-Pro)
        ('PRD-001', 'Budget Tablet 10"',         'Electronics',    179.99, 300),
        ('PRD-002', 'EcoPhone SE',               'Mobile',         299.99, 185),
        ('PRD-003', 'SwiftPen Stylus',           'Accessories',     29.99, 450),
        ('PRD-004', 'HomeHub Smart Speaker',     'Smart Home',     129.99,  95),
        # Row 6: Pro product
        ('PRD-005', 'ProMax Laptop',             'Electronics',   1299.99,  45),
        ('PRD-006', 'BasicBook Notebook',        'Stationery',       9.99, 800),
        ('PRD-007', 'SlimCase Phone Cover',      'Accessories',     14.99, 600),
        ('PRD-008', 'BrightLamp Desk',           'Lighting',        59.99, 140),
        ('PRD-009', 'QuickCharge USB-C Hub',     'Connectivity',    49.99, 260),
        ('PRD-010', 'ComfortChair Ergonomic',    'Furniture',      699.99,  22),
        # Row 12: Pro product
        ('PRD-011', 'ProSound Headphones',       'Audio',          249.99, 120),
        ('PRD-012', 'NanoRouter WiFi 6',         'Networking',     149.99,  78),
        ('PRD-013', 'ClearView Screen Protector','Accessories',     12.99, 700),
        ('PRD-014', 'SmartWatch Series 3',       'Wearables',      249.99,  90),
        ('PRD-015', 'FitBand Activity Tracker',  'Wearables',       99.99, 130),
        ('PRD-016', 'Wireless Charger Pad',      'Charging',        39.99, 350),
        ('PRD-017', 'Compact Flash Drive 128GB', 'Storage',         24.99, 500),
        # Row 19: Pro product
        ('PRD-018', 'ProEdit Camera',            'Photography',    899.99,  33),
        ('PRD-019', 'Portable Power Bank 20K',   'Charging',        59.99, 200),
        ('PRD-020', 'Gaming Headset RGB',        'Audio',          119.99,  85),
        ('PRD-021', 'Gaming Mouse 12000DPI',     'Peripherals',     79.99, 110),
        ('PRD-022', 'Streaming Microphone',      'Audio',          149.99,  65),
        ('PRD-023', 'Mini Projector HD',         'Displays',       299.99,  38),
        ('PRD-024', 'Smart Plug WiFi',           'Smart Home',      24.99, 420),
        # Row 26: Pro product
        ('PRD-025', 'ProDrive SSD 1TB',          'Storage',        189.99, 200),
        ('PRD-026', 'Air Purifier HEPA',         'Home Appliances',199.99,  50),
        ('PRD-027', 'Robot Vacuum Basic',        'Home Appliances',399.99,  35),
        ('PRD-028', 'Digital Photo Frame 10"',   'Home Decor',      79.99,  95),
        ('PRD-029', 'Bluetooth Speaker Rugged',  'Audio',           89.99, 140),
        ('PRD-030', 'Travel Adapter Universal',  'Accessories',     29.99, 500),
        ('PRD-031', 'Backlit Numpad USB',        'Peripherals',     34.99, 180),
        # Row 33: Pro product
        ('PRD-032', 'ProDesk Monitor 27"',       'Displays',       549.99,  67),
        ('PRD-033', 'Cable Management Kit',      'Organization',    19.99, 300),
        ('PRD-034', 'Monitor Cleaning Kit',      'Accessories',     14.99, 400),
        ('PRD-035', 'USB-A to USB-C Cable 2m',   'Cables',           9.99, 700),
        ('PRD-036', 'HDMI Cable 4K 2m',          'Cables',          14.99, 500),
        ('PRD-037', 'DisplayPort Cable 1.4',     'Cables',          19.99, 300),
        ('PRD-038', 'Ethernet Cable Cat6 5m',    'Cables',          12.99, 400),
        # Row 40: Pro product
        ('PRD-039', 'ProKey Mechanical Keyboard','Peripherals',    139.99,  95),
        ('PRD-040', 'Surge Protector 6-Outlet',  'Power',           39.99, 200),
        ('PRD-041', 'UPS Battery Backup 600VA',  'Power',          119.99,  55),
        ('PRD-042', 'NAS Storage 4TB',           'Storage',        349.99,  30),
        ('PRD-043', 'External DVD Drive',        'Storage',         49.99, 120),
        ('PRD-044', 'Memory Card 64GB',          'Storage',         19.99, 600),
        ('PRD-045', 'Card Reader USB-C',         'Connectivity',    22.99, 280),
        # Row 47: Pro product
        ('PRD-046', 'ProMouse Wireless',         'Peripherals',     89.99, 150),
        ('PRD-047', 'Video Capture Card',        'Video',           99.99,  70),
        ('PRD-048', 'Green Screen 150x200cm',    'Photography',     69.99,  85),
        ('PRD-049', 'Tripod Flexible 1.5m',      'Photography',     59.99, 100),
        ('PRD-050', 'Studio Lights Kit x2',      'Lighting',       149.99,  45),
        ('PRD-051', 'Lens Filter Set 52mm',      'Photography',     39.99, 130),
        ('PRD-052', 'Camera Bag Waterproof',     'Photography',     79.99,  90),
        # Row 54: Pro product
        ('PRD-053', 'ProCam Webcam 4K',          'Video',          199.99,  88),
        ('PRD-054', 'Action Camera 4K',          'Photography',    299.99,  60),
        ('PRD-055', 'Drone Mini 250g',           'Drones',         399.99,  25),
        ('PRD-056', 'VR Headset Standalone',     'Gaming',         499.99,  20),
        ('PRD-057', 'Console Controller',        'Gaming',          69.99, 150),
        ('PRD-058', 'Gaming Chair RGB',          'Furniture',      399.99,  18),
        ('PRD-059', 'LED Strip Lights 5m',       'Lighting',        29.99, 350),
        # Row 61: Pro product
        ('PRD-060', 'ProPad Drawing Tablet',     'Creative',       349.99,  42),
        ('PRD-061', 'Smart Bulb Color Pack x4',  'Smart Home',      59.99, 120),
        ('PRD-062', 'Doorbell Camera WiFi',      'Security',       149.99,  65),
        ('PRD-063', 'Security Camera Outdoor',   'Security',       179.99,  48),
        ('PRD-064', 'Smart Lock Fingerprint',    'Security',       199.99,  40),
        ('PRD-065', 'Motion Sensor PIR',         'Security',        24.99, 200),
        ('PRD-066', 'Smoke Detector Smart',      'Safety',          49.99, 150),
        # Row 68: Pro product
        ('PRD-067', 'ProStation Docking Hub',    'Connectivity',   129.99, 110),
        ('PRD-068', 'CO2 Monitor Digital',       'Safety',          79.99,  80),
        ('PRD-069', 'Thermometer Indoor/Out',    'Home Appliances', 29.99, 250),
        ('PRD-070', 'Weather Station Pro',       'Home Appliances',119.99,  55),
        ('PRD-071', 'Digital Kitchen Scale',     'Kitchen',         19.99, 300),
        ('PRD-072', 'Instant Pot 6Qt',           'Kitchen',        129.99,  70),
        ('PRD-073', 'Air Fryer 5L',              'Kitchen',         99.99,  85),
        # Row 75: Pro product
        ('PRD-074', 'ProLight Ring LED',         'Lighting',        79.99, 175),
        ('PRD-075', 'Coffee Grinder Burr',       'Kitchen',         89.99,  60),
        ('PRD-076', 'Sous Vide Cooker',          'Kitchen',         79.99,  45),
        ('PRD-077', 'Blender High Power',        'Kitchen',         79.99,  90),
        ('PRD-078', 'Electric Kettle 1.7L',      'Kitchen',         39.99, 180),
        ('PRD-079', 'Toaster 4-Slice',           'Kitchen',         49.99, 140),
        ('PRD-080', 'Microwave Compact 700W',    'Kitchen',        129.99,  55),
        # Row 82: Pro product
        ('PRD-081', 'ProAudio Mixer 8CH',        'Audio',          499.99,  28),
        ('PRD-082', 'Handheld Vacuum',           'Home Appliances', 59.99, 100),
        ('PRD-083', 'Steam Iron 2400W',          'Home Appliances', 59.99, 110),
        ('PRD-084', 'Electric Shaver Pro',       'Personal Care',   89.99,  90),
        ('PRD-085', 'Hair Dryer 2000W',          'Personal Care',   59.99, 120),
        ('PRD-086', 'Electric Toothbrush',       'Personal Care',   49.99, 150),
        ('PRD-087', 'Foot Massager Shiatsu',     'Personal Care',   79.99,  65),
        # Row 89: Pro product
        ('PRD-088', 'ProCase Laptop Bag',        'Accessories',     69.99, 220),
        ('PRD-089', 'Neck Massager Heat',        'Personal Care',   49.99,  85),
        ('PRD-090', 'Yoga Mat Premium',          'Sports',          39.99, 200),
        ('PRD-091', 'Resistance Band Set',       'Sports',          24.99, 300),
        ('PRD-092', 'Dumbbell Set 20kg',         'Sports',         149.99,  40),
        ('PRD-093', 'Foam Roller',               'Sports',          24.99, 250),
        ('PRD-094', 'Jump Rope Speed',           'Sports',          14.99, 400),
        # Row 96: Pro product
        ('PRD-095', 'ProStand Monitor Arm',      'Furniture',      119.99,  60),
        ('PRD-096', 'Heart Rate Monitor',        'Sports',          79.99,  90),
        ('PRD-097', 'Running Belt Waist',        'Sports',          19.99, 300),
        ('PRD-098', 'Bike Lock Heavy Duty',      'Sports',          29.99, 180),
        ('PRD-099', 'Camping Headlamp 300lm',    'Outdoor',         29.99, 220),
        ('PRD-100', 'Waterproof Backpack 30L',   'Outdoor',         79.99, 100),
        ('PRD-101', 'Trekking Poles Pair',       'Outdoor',         59.99,  80),
        ('PRD-102', 'Sleeping Bag -5C',          'Outdoor',         89.99,  60),
        ('PRD-103', 'Portable Stove Butane',     'Outdoor',         39.99, 120),
        ('PRD-104', 'Water Filter Bottle',       'Outdoor',         34.99, 150),
        ('PRD-105', 'Solar Panel Portable 20W',  'Outdoor',         99.99,  55),
        ('PRD-106', 'Emergency Radio Crank',     'Safety',          39.99, 100),
        ('PRD-107', 'First Aid Kit Deluxe',      'Safety',          29.99, 180),
        ('PRD-108', 'Multi-Tool Pocket',         'Outdoor',         49.99, 200),
        ('PRD-109', 'Tactical Flashlight 1000lm','Outdoor',         39.99, 160),
        ('PRD-110', 'Satellite Communicator',    'Outdoor',        449.99,  15),
        ('PRD-111', 'Snow Chains Compact',       'Automotive',      79.99,  70),
        ('PRD-112', 'Car Jump Starter 2000A',    'Automotive',     129.99,  55),
        ('PRD-113', 'Dash Cam 4K',               'Automotive',     179.99,  65),
        ('PRD-114', 'OBD2 Scanner Bluetooth',    'Automotive',      49.99, 120),
        ('PRD-115', 'Car Vacuum 12V',            'Automotive',      34.99, 140),
        ('PRD-116', 'Tire Inflator Portable',    'Automotive',      59.99, 100),
        ('PRD-117', 'Phone Mount Dashboard',     'Automotive',      19.99, 350),
        ('PRD-118', 'Seat Gap Organizer',        'Automotive',      24.99, 250),
        ('PRD-119', 'USB Car Charger 65W',       'Automotive',      29.99, 300),
        ('PRD-120', 'Ambient Interior Lights',   'Automotive',      39.99, 180),
    ]

    assert len(all_products) == 120, f"Expected 120 products, got {len(all_products)}"

    for r, (pid, name, cat, price, stock) in enumerate(all_products, 2):
        ws.cell(row=r, column=1, value=pid)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=cat)
        ws.cell(row=r, column=4, value=price)
        ws.cell(row=r, column=5, value=stock)

    # Enable AutoFilter on the header row — no filter applied
    ws.auto_filter.ref = 'A1:E121'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    pro_count = sum(1 for _, name, _, _, _ in all_products if name.startswith('Pro'))
    print(f'  Sheet: Products, {len(all_products)} data rows')
    print(f'  Pro products: {pro_count}')
    print(f'  AutoFilter: enabled on A1:E121, no active filter')


create_initial()
