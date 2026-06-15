"""
Reward Script: Filter product list so only products beginning with 'Pro' are shown
Task ID: calc_dop_filter_beginswith_013
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Rows containing non-'Pro' products are hidden (106 rows hidden)
  Component 2 (0.5): All visible rows contain products whose names start with 'Pro'
                     AND exactly 14 such rows are visible
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_filter_beginswith_013'
SHEET_NAME = 'Products'
EXPECTED_VISIBLE_COUNT = 14      # number of products starting with 'Pro'
EXPECTED_HIDDEN_COUNT = 106      # 120 total data rows - 14 Pro products
TOTAL_DATA_ROWS = 120            # rows 2-121 are data rows


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires applying a 'begins with Pro' filter on column B (Product Name).
    In xlsx format, this filter is implemented by hiding rows where the product name
    does NOT start with 'Pro'. The golden file should have:
    - 106 rows hidden (non-Pro products)
    - 14 rows visible (Pro products only)
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Precondition: Verify the sheet has the expected number of data rows
    if ws.max_row < TOTAL_DATA_ROWS + 1:
        print(f"CRITICAL: Sheet has only {ws.max_row} rows, expected at least {TOTAL_DATA_ROWS + 1}")
        print("REWARD: 0.0")
        return 0.0

    # Determine which rows are hidden and which are visible
    hidden_rows = []
    visible_data_rows = []
    for row_idx in range(2, TOTAL_DATA_ROWS + 2):  # rows 2 through 121
        rd = ws.row_dimensions.get(row_idx)
        if rd is not None and rd.hidden:
            hidden_rows.append(row_idx)
        else:
            visible_data_rows.append(row_idx)

    # Component 1: Non-Pro products are hidden (0.5 points)
    # The filter hides all rows where Product Name does NOT start with 'Pro'.
    # We expect exactly 106 rows to be hidden.
    try:
        if len(hidden_rows) == EXPECTED_HIDDEN_COUNT:
            print(f"PASS: Component 1 — {len(hidden_rows)} rows are hidden as expected "
                  f"(non-'Pro' products filtered out) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Expected {EXPECTED_HIDDEN_COUNT} hidden rows, "
                  f"found {len(hidden_rows)} hidden rows")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All visible rows have product names starting with 'Pro' AND
    # exactly 14 such rows are visible (0.5 points)
    # This verifies both that the right rows are shown and the wrong ones are hidden.
    try:
        if len(visible_data_rows) != EXPECTED_VISIBLE_COUNT:
            print(f"FAIL: Component 2 — Expected {EXPECTED_VISIBLE_COUNT} visible rows, "
                  f"found {len(visible_data_rows)} visible rows")
        else:
            # Verify each visible row has a product name starting with 'Pro'
            non_pro_visible = []
            for row_idx in visible_data_rows:
                name = ws.cell(row=row_idx, column=2).value
                if name is None or not str(name).startswith('Pro'):
                    non_pro_visible.append((row_idx, name))

            if len(non_pro_visible) == 0:
                print(f"PASS: Component 2 — All {len(visible_data_rows)} visible rows have "
                      f"product names starting with 'Pro' (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — {len(non_pro_visible)} visible rows have "
                      f"product names NOT starting with 'Pro': {non_pro_visible[:5]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
