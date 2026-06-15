"""
Reward Script: Wedding planning budget with vendor categories, payment status, and spending chart
Task ID: calc_gpm_086
Domain: libreoffice_calc
Scoring:
  Component 1: Merged title cells A1:H1 and A2:H2 with proper styling (0.15)
  Component 2: Header row 4 bold centered with fill + white font (0.10)
  Component 3: Formulas in F (Balance=D-E) and G (% of Budget=D/45000) columns (0.25)
  Component 4: Data validation on H5:H22 with payment status list (0.15)
  Component 5: Conditional formatting rules on H column (0.10)
  Component 6: Pie chart present (0.15)
  Component 7: TOTAL row 23 with SUM formulas and Remaining Budget row 24 (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_086'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Wedding' not in wb.sheetnames:
        print("FAIL: 'Wedding' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Wedding']

    # Component 1: Merged title cells A1:H1 and A2:H2 with styling (0.15 points)
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_a1_merge = any('A1' in r and 'H1' in r for r in merged_ranges)
        has_a2_merge = any('A2' in r and 'H2' in r for r in merged_ranges)

        if has_a1_merge and has_a2_merge:
            # Also check title text and styling
            a1 = ws['A1']
            title_ok = a1.value and 'Chen' in str(a1.value) and 'Rodriguez' in str(a1.value)
            bold_ok = a1.font.bold == True
            size_ok = a1.font.size is not None and a1.font.size >= 14
            centered_ok = a1.alignment.horizontal == 'center'
            fill_ok = a1.fill.fill_type == 'solid'

            a2 = ws['A2']
            italic_ok = a2.font.italic == True
            a2_fill_ok = a2.fill.fill_type == 'solid'

            sub_checks = sum([title_ok, bold_ok, size_ok, centered_ok, fill_ok, italic_ok, a2_fill_ok])
            if sub_checks >= 5:
                print(f"PASS: Component 1 — Merged titles with proper styling ({sub_checks}/7 sub-checks) (0.15 pts)")
                total_score += 0.15
            elif sub_checks >= 3:
                print(f"PARTIAL: Component 1 — Merged titles but some styling missing ({sub_checks}/7) (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 1 — Merged titles but styling insufficient ({sub_checks}/7)")
        else:
            print(f"FAIL: Component 1 — Title cells not merged. Merged ranges: {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row 4 bold centered with rose fill and white font (0.10 points)
    try:
        header_values = ['Category', 'Vendor', 'Estimate', 'Actual', 'Paid', 'Balance', '% of Budget', 'Payment Status']
        bold_count = 0
        centered_count = 0
        fill_count = 0
        white_font_count = 0

        for col_idx in range(1, 9):
            cell = ws.cell(row=4, column=col_idx)
            if cell.font.bold:
                bold_count += 1
            if cell.alignment.horizontal == 'center':
                centered_count += 1
            if cell.fill.fill_type == 'solid':
                fill_count += 1
            try:
                fc = cell.font.color.rgb
                if fc and 'FFFFFF' in str(fc).upper():
                    white_font_count += 1
            except:
                pass

        # The key differentiator from initial: bold + centered + fill + white font
        if bold_count >= 6 and centered_count >= 6 and fill_count >= 6:
            print(f"PASS: Component 2 — Headers bold({bold_count}/8) centered({centered_count}/8) filled({fill_count}/8) white_font({white_font_count}/8) (0.10 pts)")
            total_score += 0.10
        elif bold_count >= 4 and fill_count >= 4:
            print(f"PARTIAL: Component 2 — Some header styling (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — Headers not properly styled. bold={bold_count}, centered={centered_count}, fill={fill_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas in F (Balance=D-E) and G (% of Budget=D/45000) columns (0.25 points)
    try:
        f_formula_count = 0
        g_formula_count = 0

        for row in range(5, 23):  # rows 5-22 (18 items)
            f_val = ws.cell(row=row, column=6).value  # F column (Balance)
            g_val = ws.cell(row=row, column=7).value  # G column (% of Budget)

            if isinstance(f_val, str) and f_val.startswith('='):
                # Check it references D and E in same row
                if f'D{row}' in f_val and f'E{row}' in f_val:
                    f_formula_count += 1
            if isinstance(g_val, str) and g_val.startswith('='):
                # Check it references D and 45000
                if f'D{row}' in g_val and '45000' in g_val:
                    g_formula_count += 1

        f_score = min(f_formula_count / 18.0, 1.0)
        g_score = min(g_formula_count / 18.0, 1.0)
        combined = (f_score + g_score) / 2.0

        if combined >= 0.9:
            print(f"PASS: Component 3 — F formulas: {f_formula_count}/18, G formulas: {g_formula_count}/18 (0.25 pts)")
            total_score += 0.25
        elif combined >= 0.5:
            pts = round(0.25 * combined, 2)
            print(f"PARTIAL: Component 3 — F formulas: {f_formula_count}/18, G formulas: {g_formula_count}/18 ({pts} pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 3 — F formulas: {f_formula_count}/18, G formulas: {g_formula_count}/18")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Data validation on H column for payment statuses (0.15 points)
    try:
        dv_list = ws.data_validations.dataValidation
        payment_dv_found = len([
            dv for dv in dv_list
            if dv.type == 'list' and dv.formula1
            and 'paid' in str(dv.formula1).lower()
            and ('not' in str(dv.formula1).lower() or 'deposit' in str(dv.formula1).lower())
            and 'H' in str(dv.sqref).upper()
        ])

        if payment_dv_found > 0:
            print(f"PASS: Component 4 — Payment status data validation found on H column (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — No payment status data validation found. DVs: {len(dv_list)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting rules on H column (0.10 points)
    try:
        cf_rules_on_h = 0
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'H' in cf_range.upper():
                cf_rules_on_h += len(cf.rules)

        if cf_rules_on_h >= 3:
            print(f"PASS: Component 5 — {cf_rules_on_h} conditional formatting rules on H column (0.10 pts)")
            total_score += 0.10
        elif cf_rules_on_h >= 1:
            print(f"PARTIAL: Component 5 — Only {cf_rules_on_h} CF rules on H (need >= 3) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting rules on H column")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Pie chart present with Budget Allocation title (0.15 points)
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Check for pie chart
            pie_count = len([ch for ch in charts if 'Pie' in ch.__class__.__name__])

            if pie_count > 0:
                print(f"PASS: Component 6 — Pie chart found ({len(charts)} chart(s) total) (0.15 pts)")
                total_score += 0.15
            elif len(charts) > 0:
                # There's a chart, just not a pie — partial credit
                print(f"PARTIAL: Component 6 — Chart found but not a pie chart (0.08 pts)")
                total_score += 0.08
        else:
            print(f"FAIL: Component 6 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: TOTAL row 23 with SUM formulas and Remaining Budget row 24 (0.10 points)
    try:
        total_label = ws.cell(row=23, column=1).value
        remaining_label = ws.cell(row=24, column=1).value

        has_total_label = total_label and 'TOTAL' in str(total_label).upper()
        has_remaining_label = remaining_label and 'REMAINING' in str(remaining_label).upper()

        # Check SUM formulas in row 23
        sum_count = 0
        for col in range(3, 8):  # C through G
            val = ws.cell(row=23, column=col).value
            if isinstance(val, str) and '=SUM' in val.upper():
                sum_count += 1

        # Check remaining budget formula in row 24
        d24 = ws.cell(row=24, column=4).value
        has_remaining_formula = isinstance(d24, str) and '=' in d24 and '45000' in str(d24)

        # Bold on TOTAL row
        total_bold = ws.cell(row=23, column=1).font.bold == True

        # Only score task-introduced changes: SUM formulas, remaining budget formula, bold
        # Labels exist in initial too, so they are preconditions, not scoring items
        task_changes = sum([sum_count >= 3, has_remaining_formula, total_bold])

        if task_changes >= 3:
            print(f"PASS: Component 7 — TOTAL row with {sum_count} SUM formulas, Remaining Budget formula present, bold={total_bold} (0.10 pts)")
            total_score += 0.10
        elif task_changes >= 2:
            print(f"PARTIAL: Component 7 — {task_changes}/3 task-change checks passed (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — TOTAL row missing task changes. SUM count={sum_count}, remaining_formula={has_remaining_formula}, bold={total_bold}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
