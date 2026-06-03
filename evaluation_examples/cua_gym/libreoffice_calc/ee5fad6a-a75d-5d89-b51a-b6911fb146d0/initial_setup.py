"""
Initial Setup: Wedding planning budget - raw data without formatting
Task ID: calc_gpm_086
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Wedding'

    # Row 1: Title (plain text, no merge, no formatting)
    ws['A1'] = 'Wedding Budget Planner - Chen & Rodriguez'

    # Row 2: Subtitle (plain text, no merge, no formatting)
    ws['A2'] = 'Date: September 12, 2026 | Budget: $45,000 | Guests: 150'

    # Row 4: Headers (plain text, no formatting)
    headers = ['Category', 'Vendor', 'Estimate', 'Actual', 'Paid', 'Balance', '% of Budget', 'Payment Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    # 18 vendor line items with raw data (no formulas, no formatting)
    # Columns: Category, Vendor, Estimate, Actual, Paid, Balance, % of Budget, Payment Status
    items = [
        ('Venue',        'Grand Ballroom',  12000, 12000, 12000, 0,     0.2667, 'Paid in Full'),
        ('Catering',     'Elite Catering',   8500,  8500,  4250, 4250,  0.1889, 'Deposit Paid'),
        ('Photography',  'Lens Studio',      3500,  3500,  3500, 0,     0.0778, 'Paid in Full'),
        ('Videography',  'Motion Films',     2500,  2500,  1250, 1250,  0.0556, 'Deposit Paid'),
        ('Flowers',      'Bloom & Co',       2800,  2800,  1400, 1400,  0.0622, 'Deposit Paid'),
        ('Music/DJ',     'SoundWave',        1800,  1800,     0, 1800,  0.0400, 'Not Paid'),
        ('Cake',         'Sweet Layers',      900,   900,   900, 0,     0.0200, 'Paid in Full'),
        ('Dress',        'Bridal Boutique',  2200,  2200,  2200, 0,     0.0489, 'Paid in Full'),
        ('Suits',        'Formal Wear',       800,   800,   400, 400,   0.0178, 'Deposit Paid'),
        ('Hair/Makeup',  'Glam Team',         600,   600,   600, 0,     0.0133, 'Paid in Full'),
        ('Invitations',  'Paper Press',       450,   450,   450, 0,     0.0100, 'Paid in Full'),
        ('Favors',       'Gift Co',           375,   375,     0, 375,   0.0083, 'Not Paid'),
        ('Transport',    'Limo Service',      650,   650,   325, 325,   0.0144, 'Deposit Paid'),
        ('Officiant',    '',                  300,   300,   300, 0,     0.0067, 'Paid in Full'),
        ('Decor',        'Design House',     1500,  1500,     0, 1500,  0.0333, 'Not Paid'),
        ('Rentals',      'Party Rental',     1200,  1200,   600, 600,   0.0267, 'Deposit Paid'),
        ('Tips & Misc',  '',                 1000,  1000,     0, 1000,  0.0222, 'Not Paid'),
        ('Emergency Fund', '',               2000,  2000,     0, 2000,  0.0444, 'Not Paid'),
    ]

    for r, item in enumerate(items, 5):
        ws.cell(row=r, column=1, value=item[0])  # Category
        ws.cell(row=r, column=2, value=item[1])  # Vendor
        ws.cell(row=r, column=3, value=item[2])  # Estimate
        ws.cell(row=r, column=4, value=item[3])  # Actual
        ws.cell(row=r, column=5, value=item[4])  # Paid
        ws.cell(row=r, column=6, value=item[5])  # Balance (raw number)
        ws.cell(row=r, column=7, value=item[6])  # % of Budget (raw number)
        ws.cell(row=r, column=8, value=item[7])  # Payment Status

    # Row 23: TOTAL label and raw sums (no formatting)
    ws.cell(row=23, column=1, value='TOTAL')
    ws.cell(row=23, column=3, value=sum(i[2] for i in items))
    ws.cell(row=23, column=4, value=sum(i[3] for i in items))
    ws.cell(row=23, column=5, value=sum(i[4] for i in items))
    ws.cell(row=23, column=6, value=sum(i[5] for i in items))

    # Row 24: Remaining Budget (raw number, no formula, no formatting)
    ws.cell(row=24, column=1, value='Remaining Budget')
    ws.cell(row=24, column=4, value=45000 - sum(i[3] for i in items))

    # NO merges, NO formatting, NO formulas, NO chart, NO data validation,
    # NO conditional formatting - these are what the agent needs to build

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
