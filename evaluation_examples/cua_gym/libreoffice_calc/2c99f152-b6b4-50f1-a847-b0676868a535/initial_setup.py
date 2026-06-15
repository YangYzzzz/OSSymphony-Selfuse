"""
Initial Setup: IFERROR with INDEX/MATCH lookup for non-existent employee
Task ID: calc_lf_020
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Staff'

    # Headers
    ws['A1'] = 'Name'
    ws['B1'] = 'Department'
    ws['C1'] = 'Salary'

    # Employee data
    ws['A2'] = 'John'
    ws['B2'] = 'IT'
    ws['C2'] = 72000

    ws['A3'] = 'Jane'
    ws['B3'] = 'HR'
    ws['C3'] = 65000

    ws['A4'] = 'Mark'
    ws['B4'] = 'Finance'
    ws['C4'] = 80000

    # Lookup area
    ws['E1'] = 'Search Name'
    ws['E2'] = 'Zara'
    ws['F1'] = 'Salary'
    # F2 intentionally left empty — task is to enter the formula here

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
