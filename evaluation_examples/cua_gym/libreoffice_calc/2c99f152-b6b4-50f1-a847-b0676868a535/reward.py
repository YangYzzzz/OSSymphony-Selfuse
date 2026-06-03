"""
Reward Script: IFERROR with INDEX/MATCH for employee lookup
Task ID: calc_lf_020
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): F2 contains an IFERROR formula
  - Component 2 (0.3): F2 uses INDEX/MATCH pattern with correct ranges
  - Component 3 (0.3): F2 error string is 'N/A - Employee not found'
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_020'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Staff' sheet must exist
    if 'Staff' not in wb.sheetnames:
        print("FAIL: 'Staff' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Staff']
    f2_val = ws['F2'].value

    # Component 1: F2 contains an IFERROR formula (0.4 points)
    # This checks that the cell has a formula string starting with =IFERROR
    try:
        if f2_val is not None and isinstance(f2_val, str) and f2_val.strip().upper().startswith('=IFERROR('):
            print(f"PASS: Component 1 — F2 contains IFERROR formula (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Expected IFERROR formula in F2, found: {f2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2 formula uses INDEX/MATCH with correct ranges (0.3 points)
    # Verify the formula contains INDEX(...MATCH...) referencing C2:C4, A2:A4, E2
    try:
        if f2_val is not None and isinstance(f2_val, str):
            formula_upper = f2_val.upper().replace(" ", "")
            has_index = "INDEX(" in formula_upper
            has_match = "MATCH(" in formula_upper
            # Check ranges: C2:C4 for INDEX, A2:A4 and E2 for MATCH
            has_c_range = "C2:C4" in f2_val.upper().replace(" ", "")
            has_a_range = "A2:A4" in f2_val.upper().replace(" ", "")
            has_e2_ref = "E2" in f2_val.upper().replace(" ", "")

            if has_index and has_match and has_c_range and has_a_range and has_e2_ref:
                print(f"PASS: Component 2 — INDEX/MATCH with correct ranges (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not has_index:
                    missing.append("INDEX")
                if not has_match:
                    missing.append("MATCH")
                if not has_c_range:
                    missing.append("C2:C4")
                if not has_a_range:
                    missing.append("A2:A4")
                if not has_e2_ref:
                    missing.append("E2")
                print(f"FAIL: Component 2 — Missing: {', '.join(missing)} in formula: {f2_val!r}")
        else:
            print(f"FAIL: Component 2 — F2 is not a formula string: {f2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: F2 error message is 'N/A - Employee not found' (0.3 points)
    # The IFERROR second argument should be the exact error string
    try:
        if f2_val is not None and isinstance(f2_val, str):
            # Look for the error string in the formula (within quotes)
            if 'N/A - Employee not found' in f2_val or 'N/A - Employee not found' in f2_val:
                print(f"PASS: Component 3 — Error message 'N/A - Employee not found' present (0.3 pts)")
                total_score += 0.3
            else:
                # Also check case-insensitive
                if 'n/a - employee not found' in f2_val.lower():
                    print(f"PASS: Component 3 — Error message present (case-insensitive match) (0.3 pts)")
                    total_score += 0.3
                else:
                    print(f"FAIL: Component 3 — Error message 'N/A - Employee not found' not found in: {f2_val!r}")
        else:
            print(f"FAIL: Component 3 — F2 is not a formula string: {f2_val!r}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state("libreoffice_calc")

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
