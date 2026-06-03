"""
Reward Script: Personal Fitness Goal Tracking Spreadsheet
Task ID: calc_grs_035
Domain: libreoffice_calc
Scoring:
  Component 1: Monthly Summary COUNTIF/SUMIF formulas (0.25)
  Component 2: Monthly Summary overall statistics formulas (0.15)
  Component 3: Progress sheet formulas and reference data (0.15)
  Component 4: Conditional formatting on Mood Rating column (0.15)
  Component 5: Weekly Summary sheet with weekly data (0.15)
  Component 6: Line chart on Weekly Summary sheet (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_035'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Workout Log sheet must exist with data
    if 'Workout Log' not in wb.sheetnames:
        print("CRITICAL: 'Workout Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Monthly Summary sheet must exist
    if 'Monthly Summary' not in wb.sheetnames:
        print("CRITICAL: 'Monthly Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_log = wb['Workout Log']
    ws_summary = wb['Monthly Summary']

    # =========================================================================
    # Component 1: Monthly Summary COUNTIF/SUMIF formulas (0.25 points)
    # In golden: B5-D9 have COUNTIF/SUMIF formulas for each workout type.
    # In initial: B5-D9 are empty (only row labels A5-A9 exist).
    # =========================================================================
    try:
        comp1_score = 0.0
        workout_types = ['Cardio', 'Strength', 'Flexibility', 'HIIT', 'Rest']
        formulas_found = 0
        expected_formulas = 0

        for row_idx, wtype in enumerate(workout_types, start=5):
            # Check COUNTIF in column B
            b_val = ws_summary.cell(row=row_idx, column=2).value
            if b_val and isinstance(b_val, str) and 'COUNTIF' in b_val.upper():
                formulas_found += 1
            expected_formulas += 1

            # Check SUMIF in column C (Total Minutes)
            c_val = ws_summary.cell(row=row_idx, column=3).value
            if c_val and isinstance(c_val, str) and 'SUMIF' in c_val.upper():
                formulas_found += 1
            expected_formulas += 1

            # Check SUMIF in column D (Total Calories)
            d_val = ws_summary.cell(row=row_idx, column=4).value
            if d_val and isinstance(d_val, str) and 'SUMIF' in d_val.upper():
                formulas_found += 1
            expected_formulas += 1

        if expected_formulas > 0:
            ratio = formulas_found / expected_formulas
            comp1_score = 0.25 * ratio

        if comp1_score >= 0.20:
            print(f"PASS: Component 1 — COUNTIF/SUMIF formulas: {formulas_found}/{expected_formulas} found ({comp1_score:.2f} pts)")
        else:
            print(f"FAIL: Component 1 — COUNTIF/SUMIF formulas: {formulas_found}/{expected_formulas} found ({comp1_score:.2f} pts)")
        total_score += comp1_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Monthly Summary overall statistics formulas (0.15 points)
    # In golden: B12 has COUNTA, B13 has SUM, B14 has AVERAGE, B15 has COUNTA.
    # In initial: B12-B15 are empty.
    # =========================================================================
    try:
        comp2_score = 0.0
        stats_checks = 0

        # B12: Total Workouts (COUNTA or COUNT)
        b12 = ws_summary.cell(row=12, column=2).value
        if b12 and isinstance(b12, str) and ('COUNTA' in b12.upper() or 'COUNT' in b12.upper()):
            stats_checks += 1

        # B13: Total Minutes Exercised (SUM)
        b13 = ws_summary.cell(row=13, column=2).value
        if b13 and isinstance(b13, str) and 'SUM' in b13.upper():
            stats_checks += 1

        # B14: Average Session Duration (AVERAGE)
        b14 = ws_summary.cell(row=14, column=2).value
        if b14 and isinstance(b14, str) and 'AVERAGE' in b14.upper():
            stats_checks += 1

        # B15: Days Worked Out (COUNTA or COUNT)
        b15 = ws_summary.cell(row=15, column=2).value
        if b15 and isinstance(b15, str) and ('COUNTA' in b15.upper() or 'COUNT' in b15.upper()):
            stats_checks += 1

        if stats_checks >= 4:
            comp2_score = 0.15
        elif stats_checks >= 2:
            comp2_score = 0.15 * (stats_checks / 4.0)
        else:
            comp2_score = 0.0

        if comp2_score >= 0.12:
            print(f"PASS: Component 2 — Overall statistics formulas: {stats_checks}/4 found ({comp2_score:.2f} pts)")
        else:
            print(f"FAIL: Component 2 — Overall statistics formulas: {stats_checks}/4 found ({comp2_score:.2f} pts)")
        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Progress sheet formulas and reference data (0.15 points)
    # In golden: Progress sheet has previous month values (B4-B9 numeric),
    # current month formulas (C4-C9 referencing Monthly Summary), and
    # change formulas (D4-D9 = C-B).
    # In initial: Progress sheet has only row labels, no formulas/values.
    # =========================================================================
    try:
        comp3_score = 0.0
        if 'Progress' in wb.sheetnames:
            ws_prog = wb['Progress']
            formulas_or_values = 0
            total_checks = 0

            for row_idx in range(4, 10):
                # B column: previous month values (should be numeric)
                b_val = ws_prog.cell(row=row_idx, column=2).value
                if b_val is not None:
                    formulas_or_values += 1
                total_checks += 1

                # C column: current month formulas (referencing Monthly Summary)
                c_val = ws_prog.cell(row=row_idx, column=3).value
                if c_val is not None:
                    formulas_or_values += 1
                total_checks += 1

                # D column: change formulas
                d_val = ws_prog.cell(row=row_idx, column=4).value
                if d_val is not None:
                    formulas_or_values += 1
                total_checks += 1

            if total_checks > 0:
                ratio = formulas_or_values / total_checks
                comp3_score = 0.15 * ratio

            if comp3_score >= 0.12:
                print(f"PASS: Component 3 — Progress sheet data: {formulas_or_values}/{total_checks} cells populated ({comp3_score:.2f} pts)")
            else:
                print(f"FAIL: Component 3 — Progress sheet data: {formulas_or_values}/{total_checks} cells populated ({comp3_score:.2f} pts)")
        else:
            print("FAIL: Component 3 — 'Progress' sheet not found")

        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Conditional formatting on Mood Rating column (0.15 points)
    # In golden: colorScale conditional formatting on G2:G24.
    # In initial: no conditional formatting at all.
    # =========================================================================
    try:
        comp4_score = 0.0
        cf_on_mood = False

        for cf in ws_log.conditional_formatting:
            cf_range_str = str(cf)
            # Check if the conditional formatting covers column G
            if 'G' in cf_range_str:
                for rule in cf.rules:
                    # Accept colorScale or any type of conditional formatting on G
                    if rule.type in ('colorScale', 'cellIs', 'dataBar', 'iconSet'):
                        cf_on_mood = True
                        break
            if cf_on_mood:
                break

        if cf_on_mood:
            comp4_score = 0.15
            print(f"PASS: Component 4 — Conditional formatting on Mood Rating column ({comp4_score:.2f} pts)")
        else:
            print("FAIL: Component 4 — No conditional formatting found on Mood Rating (G) column")

        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Weekly Summary sheet with weekly exercise data (0.15 points)
    # In golden: 'Weekly Summary' sheet with weeks and total exercise minutes.
    # In initial: this sheet does not exist at all.
    # =========================================================================
    try:
        comp5_score = 0.0

        # Look for a sheet with weekly data (could be named differently)
        weekly_sheet = None
        for sname in wb.sheetnames:
            if 'weekly' in sname.lower() or 'week' in sname.lower():
                weekly_sheet = wb[sname]
                break

        if weekly_sheet is not None:
            # Check it has at least a few rows of data with week labels and values
            data_rows = 0
            for row in weekly_sheet.iter_rows(min_row=2, max_row=weekly_sheet.max_row, min_col=1, max_col=2):
                a_val = row[0].value
                b_val = row[1].value if len(row) > 1 else None
                if a_val is not None and b_val is not None:
                    data_rows += 1

            if data_rows >= 3:
                comp5_score = 0.15
                print(f"PASS: Component 5 — Weekly Summary sheet with {data_rows} data rows ({comp5_score:.2f} pts)")
            elif data_rows >= 1:
                comp5_score = 0.08
                print(f"PARTIAL: Component 5 — Weekly Summary sheet with only {data_rows} data rows ({comp5_score:.2f} pts)")
            else:
                print(f"FAIL: Component 5 — Weekly Summary sheet exists but has no data rows")
        else:
            print("FAIL: Component 5 — No Weekly Summary sheet found")

        total_score += comp5_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # =========================================================================
    # Component 6: Line chart on Weekly Summary sheet (0.15 points)
    # In golden: a chart exists on the Weekly Summary sheet.
    # In initial: no Weekly Summary sheet, hence no chart.
    # =========================================================================
    try:
        comp6_score = 0.0

        # Find the chart - could be on any sheet that has weekly data or a dedicated chart sheet
        chart_found = False
        for sname in wb.sheetnames:
            ws_check = wb[sname]
            if len(ws_check._charts) > 0:
                # Check if any chart is a line chart or at least exists on a weekly-related sheet
                for ch in ws_check._charts:
                    chart_found = True
                    break
            if chart_found:
                break

        if chart_found:
            comp6_score = 0.15
            print(f"PASS: Component 6 — Chart found ({comp6_score:.2f} pts)")
        else:
            print("FAIL: Component 6 — No chart found in any sheet")

        total_score += comp6_score
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
