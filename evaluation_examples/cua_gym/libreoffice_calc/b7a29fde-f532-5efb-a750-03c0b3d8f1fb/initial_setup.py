"""
Initial Setup: Personal fitness goal tracking spreadsheet
Task ID: calc_grs_035
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # =========================================================
    # Sheet 1: Workout Log
    # =========================================================
    ws_log = wb.active
    ws_log.title = "Workout Log"

    headers = ["Date", "Workout Type", "Duration (min)", "Calories Burned",
               "Distance (km)", "Notes", "Mood Rating"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws_log.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws_log.column_dimensions["A"].width = 14
    ws_log.column_dimensions["B"].width = 16
    ws_log.column_dimensions["C"].width = 16
    ws_log.column_dimensions["D"].width = 16
    ws_log.column_dimensions["E"].width = 14
    ws_log.column_dimensions["F"].width = 30
    ws_log.column_dimensions["G"].width = 14

    # Freeze header row
    ws_log.freeze_panes = "A2"

    # Realistic January workout data
    workout_data = [
        ["2025-01-02", "Cardio", 45, 380, 5.2, "Morning run in the park, felt great", 4],
        ["2025-01-03", "Strength", 60, 320, None, "Upper body focus - bench press, rows, curls", 3],
        ["2025-01-04", "Rest", None, None, None, "Active recovery - light stretching", 2],
        ["2025-01-06", "HIIT", 30, 410, None, "Tabata circuit with kettlebells", 5],
        ["2025-01-07", "Cardio", 50, 420, 6.1, "Interval training on treadmill", 4],
        ["2025-01-08", "Flexibility", 40, 150, None, "Yoga flow class at the studio", 5],
        ["2025-01-09", "Strength", 55, 340, None, "Leg day - squats, lunges, deadlifts", 3],
        ["2025-01-11", "Cardio", 35, 310, 4.0, "Easy pace jog around the neighborhood", 4],
        ["2025-01-13", "HIIT", 25, 350, None, "Bodyweight circuit - burpees, mountain climbers", 4],
        ["2025-01-14", "Strength", 65, 380, None, "Full body compound movements", 3],
        ["2025-01-15", "Flexibility", 45, 160, None, "Pilates mat class", 5],
        ["2025-01-16", "Cardio", 55, 460, 7.0, "Long run along the river trail", 5],
        ["2025-01-17", "Rest", None, None, None, "Foam rolling and mobility work", 3],
        ["2025-01-19", "Strength", 50, 310, None, "Push/pull split - shoulders and back", 4],
        ["2025-01-20", "HIIT", 28, 370, None, "Sprint intervals on stationary bike", 4],
        ["2025-01-21", "Cardio", 40, 350, 4.8, "Treadmill incline walk + jog", 3],
        ["2025-01-22", "Flexibility", 35, 130, None, "Restorative yoga session", 5],
        ["2025-01-24", "Strength", 60, 350, None, "Lower body - hip thrusts, calf raises, leg press", 4],
        ["2025-01-25", "Cardio", 50, 430, 6.5, "Outdoor run in light rain, refreshing", 4],
        ["2025-01-27", "HIIT", 30, 400, None, "CrossFit-style WOD at the box", 5],
        ["2025-01-28", "Strength", 55, 330, None, "Arms and core superset", 3],
        ["2025-01-29", "Flexibility", 40, 145, None, "Yin yoga - deep stretches", 5],
        ["2025-01-30", "Cardio", 45, 390, 5.5, "Tempo run with negative splits", 4],
    ]

    data_align = Alignment(horizontal="center", vertical="center")
    notes_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for r, row_data in enumerate(workout_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_log.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 6:  # Notes column
                cell.alignment = notes_align
            else:
                cell.alignment = data_align
            # Date format
            if c == 1 and val:
                cell.number_format = 'yyyy-mm-dd'
            # Number formats
            if c == 3 and val is not None:
                cell.number_format = '0'
            if c == 4 and val is not None:
                cell.number_format = '#,##0'
            if c == 5 and val is not None:
                cell.number_format = '0.0'

    # Data validation for Workout Type column
    dv = DataValidation(
        type="list",
        formula1='"Cardio,Strength,Flexibility,HIIT,Rest"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select a valid workout type"
    dv.errorTitle = "Invalid Workout Type"
    dv.prompt = "Select workout type"
    dv.promptTitle = "Workout Type"
    dv.add(f"B2:B100")
    ws_log.add_data_validation(dv)

    # =========================================================
    # Sheet 2: Monthly Summary (headers only, NO formulas)
    # =========================================================
    ws_summary = wb.create_sheet("Monthly Summary")

    # Title
    ws_summary.merge_cells("A1:D1")
    title_cell = ws_summary["A1"]
    title_cell.value = "Monthly Workout Summary"
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Section: Workouts by Type
    ws_summary["A3"] = "Workouts by Type"
    ws_summary["A3"].font = Font(bold=True, size=12, color="2F5496")

    type_headers = ["Workout Type", "Count", "Total Minutes", "Total Calories"]
    for col, h in enumerate(type_headers, 1):
        cell = ws_summary.cell(row=4, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
        cell.border = thin_border

    workout_types = ["Cardio", "Strength", "Flexibility", "HIIT", "Rest"]
    for r, wt in enumerate(workout_types, 5):
        ws_summary.cell(row=r, column=1, value=wt).border = thin_border
        for c in range(2, 5):
            ws_summary.cell(row=r, column=c).border = thin_border

    # Section: Overall Stats
    ws_summary["A11"] = "Overall Statistics"
    ws_summary["A11"].font = Font(bold=True, size=12, color="2F5496")

    stat_labels = ["Total Workouts", "Total Minutes Exercised",
                   "Average Session Duration (min)", "Days Worked Out"]
    for r, label in enumerate(stat_labels, 12):
        ws_summary.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws_summary.cell(row=r, column=1).border = thin_border
        ws_summary.cell(row=r, column=2).border = thin_border

    # Column widths
    ws_summary.column_dimensions["A"].width = 30
    ws_summary.column_dimensions["B"].width = 16
    ws_summary.column_dimensions["C"].width = 16
    ws_summary.column_dimensions["D"].width = 16

    # =========================================================
    # Sheet 3: Progress (structure only, NO comparison data)
    # =========================================================
    ws_progress = wb.create_sheet("Progress")

    ws_progress.merge_cells("A1:D1")
    prog_title = ws_progress["A1"]
    prog_title.value = "Monthly Progress Comparison"
    prog_title.font = Font(name="Calibri", size=14, bold=True, color="2F5496")
    prog_title.alignment = Alignment(horizontal="center", vertical="center")

    prog_headers = ["Metric", "Previous Month", "Current Month", "Change"]
    for col, h in enumerate(prog_headers, 1):
        cell = ws_progress.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
        cell.border = thin_border

    progress_metrics = ["Total Workouts", "Total Minutes", "Average Duration",
                        "Total Calories", "Cardio Sessions", "Strength Sessions"]
    for r, metric in enumerate(progress_metrics, 4):
        ws_progress.cell(row=r, column=1, value=metric).border = thin_border
        for c in range(2, 5):
            ws_progress.cell(row=r, column=c).border = thin_border

    ws_progress.column_dimensions["A"].width = 22
    ws_progress.column_dimensions["B"].width = 18
    ws_progress.column_dimensions["C"].width = 18
    ws_progress.column_dimensions["D"].width = 14

    # =========================================================
    # Save
    # =========================================================
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
