"""
Reward Script: Cross-functional team capacity planning sheet
Task ID: calc_wf_088
Domain: libreoffice_calc
Scoring:
  Component 1: Total Allocation SUM formulas in G2:G13 (0.25 pts)
  Component 2: Project FTEs Available formulas in B16:F16 (0.20 pts)
  Component 3: Capacity Gap formulas in B17:F17 (0.15 pts)
  Component 4: Conditional formatting - red on over-allocation G2:G13 (0.15 pts)
  Component 5: Conditional formatting - orange on under-resourced projects B16:F16 (0.10 pts)
  Component 6: Charts present (stacked bar + capacity vs demand) (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_088'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Capacity' sheet must exist
    if 'Capacity' not in wb.sheetnames:
        print("CRITICAL: 'Capacity' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Capacity']

    # Component 1: Total Allocation SUM formulas in G2:G13 (0.25 points)
    # In the initial file, G2:G13 are empty. The golden file has =SUM(Bx:Fx) in each.
    try:
        sum_formula_count = 0
        for row_num in range(2, 14):  # rows 2 through 13
            cell_val = ws.cell(row=row_num, column=7).value  # column G
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                expected = f"=SUM(B{row_num}:F{row_num})".upper()
                if normalized == expected:
                    sum_formula_count += 1
                # Also accept variations like =B2+C2+D2+E2+F2
                elif normalized.startswith("=") and "SUM" in normalized:
                    sum_formula_count += 1

        if sum_formula_count == 12:
            print(f"PASS: Component 1 - All 12 Total Allocation SUM formulas present (0.25 pts)")
            total_score += 0.25
        elif sum_formula_count >= 6:
            partial = 0.25 * (sum_formula_count / 12)
            print(f"PARTIAL: Component 1 - {sum_formula_count}/12 SUM formulas found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {sum_formula_count}/12 Total Allocation formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Project FTEs Available formulas in B16:F16 (0.20 points)
    # Initial file has B16:F16 empty. Golden has =SUM(col)/100 formulas.
    try:
        fte_formula_count = 0
        for col_num in range(2, 7):  # columns B through F
            cell_val = ws.cell(row=16, column=col_num).value
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                # Check for =SUM(Bx:Bx)/100 pattern
                col_letter = chr(64 + col_num)  # B=66, etc.
                expected = f"=SUM({col_letter}2:{col_letter}13)/100".upper()
                if normalized == expected:
                    fte_formula_count += 1
                # Accept any formula that references the column and divides by 100
                elif normalized.startswith("=") and "/100" in normalized:
                    fte_formula_count += 1
                elif normalized.startswith("=") and "*0.01" in normalized:
                    fte_formula_count += 1

        if fte_formula_count == 5:
            print(f"PASS: Component 2 - All 5 Project FTEs Available formulas present (0.20 pts)")
            total_score += 0.20
        elif fte_formula_count >= 1:
            partial = 0.20 * (fte_formula_count / 5)
            print(f"PARTIAL: Component 2 - {fte_formula_count}/5 FTE formulas found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No Project FTEs Available formulas found in B16:F16")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Capacity Gap formulas in B17:F17 (0.15 points)
    # Initial file has B17:F17 empty. Golden has =Bx16-Bx15 formulas.
    try:
        gap_formula_count = 0
        for col_num in range(2, 7):  # columns B through F
            cell_val = ws.cell(row=17, column=col_num).value
            if cell_val is not None and isinstance(cell_val, str):
                normalized = cell_val.upper().replace(" ", "")
                col_letter = chr(64 + col_num)
                expected = f"={col_letter}16-{col_letter}15".upper()
                if normalized == expected:
                    gap_formula_count += 1
                # Accept reversed subtraction or other formula patterns referencing row 16 and 15
                elif normalized.startswith("=") and "16" in normalized and "15" in normalized:
                    gap_formula_count += 1

        if gap_formula_count == 5:
            print(f"PASS: Component 3 - All 5 Capacity Gap formulas present (0.15 pts)")
            total_score += 0.15
        elif gap_formula_count >= 1:
            partial = 0.15 * (gap_formula_count / 5)
            print(f"PARTIAL: Component 3 - {gap_formula_count}/5 gap formulas found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No Capacity Gap formulas found in B17:F17")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Conditional formatting - red fill on over-allocated individuals G2:G13 (0.15 points)
    # Initial file has 0 conditional formatting rules. Golden has red on G2:G13 for >100%.
    try:
        cf_rules = list(ws.conditional_formatting)
        red_over_alloc_found = False
        for cf in cf_rules:
            cf_range = str(cf)
            # Check if the range covers column G (Total Allocation)
            if 'G' in cf_range:
                for rule in cf.rules:
                    # Look for cellIs > 100 or expression-based rule checking >100
                    if rule.type == 'cellIs' and rule.operator == 'greaterThan':
                        if rule.formula and any('100' in str(f) for f in rule.formula):
                            red_over_alloc_found = True
                            break
                    elif rule.type == 'expression':
                        if rule.formula and any('100' in str(f) for f in rule.formula):
                            red_over_alloc_found = True
                            break
                if red_over_alloc_found:
                    break

        if red_over_alloc_found:
            print(f"PASS: Component 4 - Conditional formatting for over-allocation on column G (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 - No conditional formatting found for over-allocation (>100%) on column G")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting - orange on under-resourced projects B16:F16 (0.10 points)
    # Initial file has 0 conditional formatting rules. Golden has orange on B16:F16 when FTEs < demand.
    try:
        cf_rules = list(ws.conditional_formatting)
        under_resource_count = 0
        for cf in cf_rules:
            cf_range = str(cf)
            for rule in cf.rules:
                # Look for expression rules comparing row 16 to row 15
                if rule.type == 'expression' and rule.formula:
                    formula_str = str(rule.formula[0]).upper()
                    if '16' in formula_str and '15' in formula_str and '<' in formula_str:
                        under_resource_count += 1

        if under_resource_count >= 3:
            print(f"PASS: Component 5 - Conditional formatting for under-resourced projects ({under_resource_count} rules) (0.10 pts)")
            total_score += 0.10
        elif under_resource_count >= 1:
            partial = 0.10 * (under_resource_count / 5)
            print(f"PARTIAL: Component 5 - {under_resource_count}/5 under-resource CF rules found ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 - No conditional formatting found for under-resourced projects on row 16")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Charts present - stacked bar chart + capacity vs demand chart (0.15 points)
    # Initial file has 0 charts. Golden has 2 charts.
    try:
        chart_count = len(ws._charts)
        if chart_count >= 2:
            print(f"PASS: Component 6 - {chart_count} charts found (0.15 pts)")
            total_score += 0.15
        elif chart_count == 1:
            print(f"PARTIAL: Component 6 - Only 1 chart found, expected at least 2 (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 6 - No charts found, expected at least 2")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
