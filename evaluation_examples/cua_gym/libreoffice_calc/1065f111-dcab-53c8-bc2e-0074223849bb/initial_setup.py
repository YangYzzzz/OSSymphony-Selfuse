"""
Initial Setup: Cross-functional team capacity planning sheet
Task ID: calc_wf_088
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_088'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Capacity'

    # --- Header row ---
    headers = ['Team Member', 'P1', 'P2', 'P3', 'P4', 'P5', 'Total Allocation']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    # --- 12 team members with allocation percentages ---
    # Allocations are raw numbers (percentages), some intentionally over 100% total
    team_data = [
        # Name,          P1,  P2,  P3,  P4,  P5
        ['Sarah Chen',       40,  20,  15,  10,   5],
        ['Marcus Johnson',   25,  30,  20,  10,  15],
        ['Priya Patel',      50,  10,   0,  20,  10],
        ['David Kim',        10,  25,  35,  15,  15],
        ['Elena Rodriguez',  30,  30,  25,  10,  15],  # total 110 - over-allocated
        ['James Wilson',     20,  15,  30,  20,  10],
        ['Aisha Mohammed',   15,  40,  10,  25,   5],
        ['Thomas Berg',      35,   5,  20,  10,  20],
        ['Maria Santos',     10,  20,  25,  30,  20],  # total 105 - over-allocated
        ['Kenji Nakamura',   20,  10,  30,  15,  25],
        ['Lisa O\'Brien',    25,  25,  10,  20,  15],
        ['Carlos Mendez',    30,  15,  20,  25,  10],
    ]

    data_align = Alignment(horizontal='center', vertical='center')
    name_align = Alignment(horizontal='left', vertical='center')

    for r, row_data in enumerate(team_data, 2):
        # Name column
        cell = ws.cell(row=r, column=1, value=row_data[0])
        cell.font = Font(name='Calibri', size=11)
        cell.alignment = name_align
        cell.border = border

        # Allocation columns (P1-P5)
        for c, val in enumerate(row_data[1:], 2):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = data_align
            cell.border = border
            cell.number_format = '0"%"'

        # Total Allocation column (G) - leave EMPTY for the task
        cell = ws.cell(row=r, column=7)
        cell.border = border
        cell.alignment = data_align

    # --- Row 14: empty separator ---

    # --- Row 15: Project Demand (FTEs) ---
    demand_label_row = 15
    ws.cell(row=demand_label_row, column=1, value='Project Demand (FTEs)').font = Font(bold=True, name='Calibri', size=11)
    ws.cell(row=demand_label_row, column=1).border = border
    ws.cell(row=demand_label_row, column=1).alignment = name_align

    demand_values = [2.8, 2.0, 1.8, 1.5, 1.2]
    for c, val in enumerate(demand_values, 2):
        cell = ws.cell(row=demand_label_row, column=c, value=val)
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.alignment = data_align
        cell.border = border
        cell.number_format = '0.0'

    # --- Row 16: Project FTEs Available - EMPTY for task ---
    ws.cell(row=16, column=1, value='Project FTEs Available').font = Font(bold=True, name='Calibri', size=11)
    ws.cell(row=16, column=1).border = border
    ws.cell(row=16, column=1).alignment = name_align
    for c in range(2, 7):
        cell = ws.cell(row=16, column=c)
        cell.border = border
        cell.alignment = data_align

    # --- Row 17: Capacity Gap - EMPTY for task ---
    ws.cell(row=17, column=1, value='Capacity Gap').font = Font(bold=True, name='Calibri', size=11)
    ws.cell(row=17, column=1).border = border
    ws.cell(row=17, column=1).alignment = name_align
    for c in range(2, 7):
        cell = ws.cell(row=17, column=c)
        cell.border = border
        cell.alignment = data_align

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 12
    ws.column_dimensions['G'].width = 18

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
