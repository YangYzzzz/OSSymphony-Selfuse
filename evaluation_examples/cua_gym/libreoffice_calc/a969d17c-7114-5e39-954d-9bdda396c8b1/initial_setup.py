"""
Initial Setup: Tutoring Center Usage Statistics
Task ID: calc_edu_tutoring_center_stats_059
Domain: libreoffice_calc

Creates a TutoringLog sheet with 400 tutoring session records.
Columns: Session ID, Student ID, Subject, Date, Hour, Tutor, Pre Grade, Post Grade
"""

import openpyxl
import random
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_tutoring_center_stats_059'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

random.seed(42)

SUBJECTS = ['Mathematics', 'English', 'Chemistry', 'Biology', 'Physics',
            'History', 'Economics', 'Computer Science']

TUTORS = [
    'Dr. Amanda Foster', 'Prof. James Nguyen', 'Ms. Rachel Kim',
    'Mr. David Okafor', 'Dr. Linda Patel', 'Mr. Carlos Rivera',
    'Ms. Sophie Brennan', 'Dr. Thomas Wu'
]

STUDENT_IDS = [f'STU{str(i).zfill(4)}' for i in range(1, 81)]  # 80 students

def random_date(start, end):
    delta = end - start
    return start + timedelta(days=random.randint(0, delta.days))

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: TutoringLog ---
    ws = wb.active
    ws.title = 'TutoringLog'

    headers = ['Session ID', 'Student ID', 'Subject', 'Date', 'Hour',
               'Tutor', 'Pre Grade', 'Post Grade']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    start_date = date(2024, 9, 2)
    end_date = date(2025, 4, 30)

    for i in range(1, 401):
        session_id = f'SES{str(i).zfill(4)}'
        student_id = random.choice(STUDENT_IDS)
        subject = random.choice(SUBJECTS)
        session_date = random_date(start_date, end_date)
        hour = random.randint(8, 17)
        tutor = random.choice(TUTORS)
        pre_grade = round(random.uniform(45, 82), 1)
        post_grade = round(pre_grade + random.uniform(3, 18), 1)
        if post_grade > 100:
            post_grade = 100.0

        row_data = [session_id, student_id, subject, session_date, hour,
                    tutor, pre_grade, post_grade]
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i + 1, column=col, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  TutoringLog: 400 rows of tutoring session data')

create_initial()
