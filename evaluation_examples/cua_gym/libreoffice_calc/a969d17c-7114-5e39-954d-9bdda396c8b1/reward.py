"""
Reward Script: Analyze tutoring center usage data and create summary report
Task ID: calc_edu_tutoring_center_stats_059
Domain: libreoffice_calc
Scoring:
  Component 1: TutoringStats sheet exists (0.2 pts)
  Component 2: Sessions per Student data present and correct (0.2 pts)
  Component 3: Sessions by Subject present, sorted descending (0.2 pts)
  Component 4: Sessions by Hour present with peak hour identified (0.2 pts)
  Component 5: Grade improvement analysis present with correct values (0.1 pts)
  Component 6: Bar chart with title 'Tutoring Sessions by Subject' (0.1 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_tutoring_center_stats_059'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: TutoringLog sheet must exist
    if 'TutoringLog' not in wb.sheetnames:
        print("CRITICAL: TutoringLog sheet missing — file is corrupted")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: TutoringStats sheet created (0.2 points)
    # The initial file has only TutoringLog; golden file adds TutoringStats
    try:
        if 'TutoringStats' in wb.sheetnames:
            ws_stats = wb['TutoringStats']
            print("PASS: Component 1 — TutoringStats sheet exists (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — TutoringStats sheet not found")
            # No TutoringStats means nothing else can be verified
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Sessions per Student section exists with correct data (0.2 points)
    # Expected: row 1 col A = 'Sessions per Student', row 2 col A = 'Student ID',
    # row 2 col B = 'Session Count', then rows with STU#### IDs and counts
    try:
        header_a1 = ws_stats.cell(row=1, column=1).value
        header_a2 = ws_stats.cell(row=2, column=1).value
        header_b2 = ws_stats.cell(row=2, column=2).value

        section_header_ok = (
            header_a1 is not None and 'Session' in str(header_a1) and 'Student' in str(header_a1)
        )
        col_headers_ok = (
            header_a2 is not None and 'Student' in str(header_a2) and
            header_b2 is not None and 'Session' in str(header_b2)
        )

        # Check that at least some student rows exist with numeric counts
        student_rows_ok = False
        for row in range(3, 20):
            stu_id = ws_stats.cell(row=row, column=1).value
            count = ws_stats.cell(row=row, column=2).value
            if stu_id and str(stu_id).startswith('STU') and isinstance(count, (int, float)) and count > 0:
                student_rows_ok = True
                break

        # Verify specific known values: STU0001 has 6 sessions, STU0049 has 12 sessions
        stu_data = {}
        for row in range(3, ws_stats.max_row + 1):
            stu_id = ws_stats.cell(row=row, column=1).value
            count = ws_stats.cell(row=row, column=2).value
            if stu_id and str(stu_id).startswith('STU'):
                stu_data[str(stu_id)] = count

        known_values_ok = (
            stu_data.get('STU0001') == 6 and
            stu_data.get('STU0049') == 12 and
            len(stu_data) >= 10  # at least 10 students listed
        )

        if section_header_ok and col_headers_ok and student_rows_ok and known_values_ok:
            print(f"PASS: Component 2 — Sessions per Student section present with correct data ({len(stu_data)} students, STU0001=6, STU0049=12) (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — Sessions per Student section issues:")
            print(f"  section_header_ok={section_header_ok} (A1='{header_a1}')")
            print(f"  col_headers_ok={col_headers_ok} (A2='{header_a2}', B2='{header_b2}')")
            print(f"  student_rows_ok={student_rows_ok}")
            print(f"  known_values_ok={known_values_ok} (STU0001={stu_data.get('STU0001')}, STU0049={stu_data.get('STU0049')}, count={len(stu_data)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Sessions by Subject section, sorted descending (0.2 points)
    # Expected: row 1 col D = 'Sessions by Subject', col D rows with subjects sorted by count desc
    # English=57, Physics=55, Computer Science=54, History=54, Economics=50, Biology=49, Mathematics=42, Chemistry=39
    try:
        header_d1 = ws_stats.cell(row=1, column=4).value
        header_d2 = ws_stats.cell(row=2, column=4).value
        header_e2 = ws_stats.cell(row=2, column=5).value

        section_header_ok = (
            header_d1 is not None and 'Subject' in str(header_d1)
        )
        col_headers_ok = (
            header_d2 is not None and 'Subject' in str(header_d2) and
            header_e2 is not None and 'Session' in str(header_e2)
        )

        # Collect subject data
        subject_data = {}
        for row in range(3, ws_stats.max_row + 1):
            subj = ws_stats.cell(row=row, column=4).value
            count = ws_stats.cell(row=row, column=5).value
            if subj and isinstance(count, (int, float)):
                subject_data[str(subj)] = count

        # Verify known values
        known_values_ok = (
            subject_data.get('English') == 57 and
            subject_data.get('Chemistry') == 39 and
            len(subject_data) >= 8  # all 8 subjects
        )

        # Check sorted descending
        counts = list(subject_data.values())
        sorted_ok = counts == sorted(counts, reverse=True)

        # Top subject should be English with 57
        top_subject_ok = False
        if subject_data:
            top_subj = ws_stats.cell(row=3, column=4).value
            top_count = ws_stats.cell(row=3, column=5).value
            top_subject_ok = (top_subj == 'English' and top_count == 57)

        if section_header_ok and col_headers_ok and known_values_ok and sorted_ok and top_subject_ok:
            print(f"PASS: Component 3 — Sessions by Subject section present, sorted descending, English=57 (top), Chemistry=39 (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Sessions by Subject section issues:")
            print(f"  section_header_ok={section_header_ok} (D1='{header_d1}')")
            print(f"  col_headers_ok={col_headers_ok}")
            print(f"  known_values_ok={known_values_ok} (English={subject_data.get('English')}, Chemistry={subject_data.get('Chemistry')}, subjects={len(subject_data)})")
            print(f"  sorted_ok={sorted_ok} (counts={counts})")
            print(f"  top_subject_ok={top_subject_ok}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Sessions by Hour section with peak hour identified as 15 (0.2 points)
    # Expected: row 1 col G = 'Sessions by Hour', col G hours 8-17, col H counts, col I peak marker
    # Hour 15 has 53 sessions (most), marked as 'Peak'
    try:
        header_g1 = ws_stats.cell(row=1, column=7).value
        header_g2 = ws_stats.cell(row=2, column=7).value
        header_h2 = ws_stats.cell(row=2, column=8).value

        section_header_ok = (
            header_g1 is not None and 'Hour' in str(header_g1)
        )
        col_headers_ok = (
            header_g2 is not None and 'Hour' in str(header_g2) and
            header_h2 is not None and 'Session' in str(header_h2)
        )

        # Collect hour data
        hour_data = {}
        peak_marker_found = False
        for row in range(3, ws_stats.max_row + 1):
            hour = ws_stats.cell(row=row, column=7).value
            count = ws_stats.cell(row=row, column=8).value
            peak_marker = ws_stats.cell(row=row, column=9).value
            if hour is not None and isinstance(count, (int, float)):
                hour_data[int(hour)] = count
                if peak_marker and 'Peak' in str(peak_marker):
                    peak_hour_in_row = int(hour)
                    peak_marker_found = True

        # Verify known values: hour 15 = 53, hour 10 = 45, 10 hours total (8-17)
        known_values_ok = (
            hour_data.get(15) == 53 and
            hour_data.get(10) == 45 and
            len(hour_data) >= 10  # all 10 hours present
        )

        # Peak hour should be 15 (highest count)
        peak_ok = False
        if hour_data:
            actual_peak = max(hour_data, key=hour_data.get)
            peak_ok = (actual_peak == 15 and peak_marker_found and peak_hour_in_row == 15)

        if section_header_ok and col_headers_ok and known_values_ok and peak_ok:
            print(f"PASS: Component 4 — Sessions by Hour present, hour 15 marked as peak with 53 sessions (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Sessions by Hour section issues:")
            print(f"  section_header_ok={section_header_ok} (G1='{header_g1}')")
            print(f"  col_headers_ok={col_headers_ok}")
            print(f"  known_values_ok={known_values_ok} (hour15={hour_data.get(15)}, hour10={hour_data.get(10)}, hours={len(hour_data)})")
            print(f"  peak_ok={peak_ok} (marker_found={peak_marker_found})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grade improvement analysis with correct group values (0.1 points)
    # Expected in rows 84-87:
    #   Row 84: 'Grade Improvement Analysis' (section header)
    #   Row 85: 'Group', 'Avg Pre Grade', 'Avg Post Grade', 'Improvement'
    #   Row 86: '3+ Sessions', 62.92, 73.42, 10.5
    #   Row 87: '< 3 Sessions', 63.03, 74.08, 11.05
    # Note: rows may vary slightly; we search for the data
    try:
        grade_section_found = False
        high_sessions_row = None
        low_sessions_row = None

        for row in range(1, ws_stats.max_row + 1):
            a_val = ws_stats.cell(row=row, column=1).value
            if a_val and 'Grade' in str(a_val) and 'Analysis' in str(a_val):
                grade_section_found = True
            if a_val and '3+' in str(a_val):
                high_sessions_row = row
            if a_val and '< 3' in str(a_val):
                low_sessions_row = row

        high_group_ok = False
        low_group_ok = False

        if high_sessions_row:
            pre_g = ws_stats.cell(row=high_sessions_row, column=2).value
            post_g = ws_stats.cell(row=high_sessions_row, column=3).value
            improve = ws_stats.cell(row=high_sessions_row, column=4).value
            if (pre_g is not None and abs(float(pre_g) - 62.92) < 1.0 and
                    post_g is not None and abs(float(post_g) - 73.42) < 1.0):
                high_group_ok = True

        if low_sessions_row:
            pre_g = ws_stats.cell(row=low_sessions_row, column=2).value
            post_g = ws_stats.cell(row=low_sessions_row, column=3).value
            if (pre_g is not None and abs(float(pre_g) - 63.03) < 1.0 and
                    post_g is not None and abs(float(post_g) - 74.08) < 1.0):
                low_group_ok = True

        if grade_section_found and high_group_ok and low_group_ok:
            print(f"PASS: Component 5 — Grade improvement analysis present with correct values (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 5 — Grade improvement analysis issues:")
            print(f"  grade_section_found={grade_section_found}")
            print(f"  high_sessions_row={high_sessions_row}, high_group_ok={high_group_ok}")
            print(f"  low_sessions_row={low_sessions_row}, low_group_ok={low_group_ok}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Bar chart on TutoringStats with title 'Tutoring Sessions by Subject' (0.1 points)
    try:
        charts = ws_stats._charts
        chart_ok = False
        chart_title_ok = False

        if len(charts) >= 1:
            chart_ok = True
            for chart in charts:
                # Check if it's a BarChart
                if hasattr(chart, 'type') or type(chart).__name__ == 'BarChart':
                    # Try to extract title text
                    try:
                        title_text = chart.title.tx.rich.p[0].r[0].t
                        if 'Tutoring Sessions by Subject' in str(title_text):
                            chart_title_ok = True
                            break
                    except (AttributeError, IndexError, TypeError):
                        # Try alternate title access
                        try:
                            if hasattr(chart.title, 'tx'):
                                # Walk through the rich text structure
                                for para in chart.title.tx.rich.p:
                                    for run in para.r:
                                        if 'Tutoring Sessions by Subject' in str(run.t):
                                            chart_title_ok = True
                                            break
                        except Exception:
                            pass

        if chart_ok and chart_title_ok:
            print(f"PASS: Component 6 — Bar chart exists on TutoringStats with title 'Tutoring Sessions by Subject' (0.1 pts)")
            total_score += 0.1
        elif chart_ok:
            print(f"FAIL: Component 6 — Chart exists but title 'Tutoring Sessions by Subject' not confirmed (chart_title_ok={chart_title_ok})")
        else:
            print(f"FAIL: Component 6 — No charts found on TutoringStats sheet (found {len(charts)} charts)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
