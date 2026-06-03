"""
Reward Script: Wildcard VLOOKUP formula in B2:B11 to match partial product names
Task ID: calc_fma_vlookup_partial_049
Domain: libreoffice_calc
Scoring:
  Component 1: At least 1 wildcard VLOOKUP formula present in B2:B11 (0.3 pts)
  Component 2: All 10 cells B2:B11 contain VLOOKUP-based formulas (0.4 pts)
  Component 3: All formulas correctly reference catalog range $A$14:$B$23 with col=2, exact match (0.3 pts)
"""

import os
import re

import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_vlookup_partial_049'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Confirm sheet 'PriceCheck' exists (precondition gate, not scored)
    if 'PriceCheck' not in wb.sheetnames:
        print("FAIL: Sheet 'PriceCheck' not found. Cannot evaluate.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['PriceCheck']

    # Component 1: At least one wildcard VLOOKUP formula present in B2:B11 (0.3 points)
    # Checks that the agent started adding the required wildcard VLOOKUP formulas.
    # A wildcard VLOOKUP must contain both "VLOOKUP" and wildcard concatenation pattern.
    try:
        wildcard_vlookup_count = 0
        for row in range(2, 12):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val and isinstance(cell_val, str):
                upper_val = cell_val.upper()
                # Must be a VLOOKUP formula with wildcard concatenation: "*"&...&"*"
                if 'VLOOKUP' in upper_val and ('\"*\"' in cell_val or '"*"' in cell_val):
                    wildcard_vlookup_count += 1

        if wildcard_vlookup_count >= 1:
            print(f"PASS: Component 1 — {wildcard_vlookup_count} wildcard VLOOKUP formula(s) found in B2:B11 (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — No wildcard VLOOKUP formulas found in B2:B11 (found {wildcard_vlookup_count})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 10 cells B2:B11 contain VLOOKUP-based formulas (0.4 points)
    # Checks full completeness — all 10 price lookup cells must be filled.
    try:
        vlookup_count = 0
        missing_cells = []
        for row in range(2, 12):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val and isinstance(cell_val, str) and 'VLOOKUP' in cell_val.upper():
                vlookup_count += 1
            else:
                missing_cells.append(f'B{row}')

        if vlookup_count == 10:
            print("PASS: Component 2 — All 10 cells B2:B11 contain VLOOKUP formulas (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — Only {vlookup_count}/10 cells have VLOOKUP. Missing: {missing_cells}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas correctly reference catalog $A$14:$B$23, col=2, exact match (0.3 points)
    # Checks that the VLOOKUP formulas use the correct lookup table and return column.
    # Also verifies the wildcard concatenation pattern uses the cell reference (e.g., A2).
    try:
        correct_formula_count = 0
        incorrect_formulas = []
        for row in range(2, 12):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val and isinstance(cell_val, str):
                upper_val = cell_val.upper()
                # Must reference catalog range $A$14:$B$23 (with or without $ on rows)
                # Accept variations: $A$14:$B$23, A14:B23, $A14:$B23, etc.
                has_catalog_ref = bool(
                    re.search(r'\$?A\$?14:\$?B\$?23', cell_val, re.IGNORECASE)
                )
                # Must use column 2 (second column of lookup range)
                # Pattern: VLOOKUP(...,2,0) or VLOOKUP(...,2,FALSE)
                has_col2 = bool(
                    re.search(r'VLOOKUP\s*\(.*?,\s*2\s*,\s*(0|FALSE)\s*\)', cell_val, re.IGNORECASE)
                )
                # Must use wildcard concatenation with row-specific cell reference
                # Pattern: "*"&A{row}&"*" or similar
                has_wildcard_ref = bool(
                    re.search(rf'["\u201c]\s*\*\s*["\u201d]\s*&\s*\$?A\$?{row}\s*&\s*["\u201c]\s*\*\s*["\u201d]', cell_val, re.IGNORECASE)
                )

                if has_catalog_ref and has_col2 and has_wildcard_ref:
                    correct_formula_count += 1
                else:
                    details = []
                    if not has_catalog_ref:
                        details.append("missing catalog ref $A$14:$B$23")
                    if not has_col2:
                        details.append("missing col=2, exact_match=0")
                    if not has_wildcard_ref:
                        details.append(f"missing wildcard ref \"*\"&A{row}&\"*\"")
                    incorrect_formulas.append(f'B{row}: {details}')

        if correct_formula_count == 10:
            print("PASS: Component 3 — All 10 formulas use correct catalog range, col=2, wildcard ref (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — Only {correct_formula_count}/10 formulas are fully correct")
            for info in incorrect_formulas[:5]:  # show up to 5 examples
                print(f"  {info}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
