"""
Initial Setup: Wildcard VLOOKUP for partial product name matching
Task ID: calc_fma_vlookup_partial_049
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_vlookup_partial_049'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'PriceCheck'

    # --- Section 1: Search Terms (Rows 1-11) ---
    # Header row
    ws['A1'] = 'Search Term'
    ws['B1'] = 'Price'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # Partial search terms in A2:A11
    search_terms = [
        'Widget',
        'Gadget',
        'Gizmo',
        'Doohickey',
        'Thingamajig',
        'Widget',
        'Gizmo',
        'Gadget',
        'Doohickey',
        'Thingamajig',
    ]
    for i, term in enumerate(search_terms, 2):
        ws.cell(row=i, column=1, value=term)
        # Column B (B2:B11) intentionally left empty — agent must fill with VLOOKUP

    # --- Row 12 is empty as separator ---

    # --- Row 13: Catalog header ---
    ws['A13'] = 'Product Name'
    ws['B13'] = 'List Price'
    ws['A13'].font = Font(bold=True)
    ws['B13'].font = Font(bold=True)

    # --- Rows 14-23: Full product catalog ---
    catalog = [
        ('Widget Pro X',         29.99),
        ('Gadget Ultra 2000',   149.99),
        ('Gizmo Deluxe',         49.99),
        ('Doohickey Standard',   12.50),
        ('Thingamajig Premium',  89.99),
        ('Widget Lite',          14.99),
        ('Gizmo Basic',          24.99),
        ('Gadget Mini',          79.99),
        ('Doohickey Plus',       19.99),
        ('Thingamajig Budget',   39.99),
    ]
    for i, (name, price) in enumerate(catalog, 14):
        ws.cell(row=i, column=1, value=name)
        ws.cell(row=i, column=2, value=price)

    # Column widths for readability
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
