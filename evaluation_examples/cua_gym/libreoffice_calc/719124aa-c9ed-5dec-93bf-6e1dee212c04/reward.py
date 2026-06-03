"""
Reward Script: Calculate tuition and fees for international students.
Task ID: calc_edu_international_student_fees_063
Domain: libreoffice_calc

Scoring:
  - Component 1: Column E (Int'l Tuition = D*2.5) formula in all 30 rows         0.20 pts
  - Component 2: Column F (Int'l Fee = 1500) formula in all 30 rows               0.15 pts
  - Component 3: Column G (Total USD = E+F) formula in all 30 rows                0.15 pts
  - Component 4: Columns H and I (VLOOKUP for currency code and rate)             0.15 pts
  - Component 5: Column J (Total Home Currency = G*I) formula in all 30 rows      0.15 pts
  - Component 6: Column L (Discrepancy Flag IF/ABS formula) in all 30 rows        0.10 pts
  - Component 7: Number formatting on G (USD currency) and J columns              0.10 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_international_student_fees_063'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheet exists
    if 'IntlStudents' not in wb.sheetnames:
        print("CRITICAL: 'IntlStudents' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['IntlStudents']

    # Component 1: Column E — Int'l Tuition = D*2.5 (0.20 points)
    # Formula should be =Drow*2.5 for rows 2-31
    # This FAILS on initial (all None) and PASSES on golden
    try:
        e_count = 0
        e_formula_ok = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=5).value  # Column E
            if val is not None:
                e_count += 1
                val_str = str(val).strip().upper().replace(' ', '')
                # Match pattern like =D2*2.5, =D3*2.5, etc.
                expected_pattern = f'=D{row}*2.5'
                if val_str == expected_pattern.upper():
                    e_formula_ok += 1
        if e_formula_ok == 30:
            print(f"PASS: Component 1 — All 30 Int'l Tuition formulas (=Drow*2.5) present and correct ({e_formula_ok}/30)")
            total_score += 0.20
        elif e_count == 30 and e_formula_ok < 30:
            print(f"PARTIAL: Component 1 — 30 cells filled but only {e_formula_ok}/30 match =Drow*2.5 pattern")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Only {e_count}/30 Int'l Tuition cells filled (need formula =Drow*2.5)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column F — Int'l Fee = 1500 (0.15 points)
    # Formula =1500 or literal value 1500 for all 30 rows
    try:
        f_count = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=6).value  # Column F
            if val is not None:
                val_str = str(val).strip().replace(' ', '')
                # Accept =1500 (formula), 1500 (numeric), or "1500"
                if val_str in ('=1500', '1500') or val == 1500:
                    f_count += 1
        if f_count == 30:
            print(f"PASS: Component 2 — All 30 Int'l Fee values (=1500) present ({f_count}/30)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — Only {f_count}/30 Int'l Fee cells have correct value/formula (need =1500 or 1500)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Column G — Total USD = E+F (0.15 points)
    # Formula should be =Erow+Frow for rows 2-31
    try:
        g_count = 0
        g_formula_ok = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=7).value  # Column G
            if val is not None:
                g_count += 1
                val_str = str(val).strip().upper().replace(' ', '')
                expected = f'=E{row}+F{row}'
                if val_str == expected.upper():
                    g_formula_ok += 1
        if g_formula_ok == 30:
            print(f"PASS: Component 3 — All 30 Total USD formulas (=Erow+Frow) present and correct ({g_formula_ok}/30)")
            total_score += 0.15
        elif g_count == 30:
            print(f"PARTIAL: Component 3 — 30 cells filled but only {g_formula_ok}/30 match =Erow+Frow pattern")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — Only {g_count}/30 Total USD cells filled (need formula =Erow+Frow)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Columns H (Home Currency code) and I (Exchange Rate) via VLOOKUP (0.15 points)
    # H should reference ExchangeRates column 2, I should reference ExchangeRates column 3
    try:
        h_count = 0
        i_count = 0
        h_vlookup_ok = 0
        i_vlookup_ok = 0
        for row in range(2, 32):
            h_val = ws.cell(row=row, column=8).value   # Column H
            i_val = ws.cell(row=row, column=9).value   # Column I
            if h_val is not None:
                h_count += 1
                h_str = str(h_val).strip().upper().replace(' ', '')
                # Check for VLOOKUP pattern referencing ExchangeRates and col_index 2
                if 'VLOOKUP' in h_str and 'EXCHANGERATES' in h_str and ',2,' in h_str:
                    h_vlookup_ok += 1
            if i_val is not None:
                i_count += 1
                i_str = str(i_val).strip().upper().replace(' ', '')
                # Check for VLOOKUP pattern referencing ExchangeRates and col_index 3
                if 'VLOOKUP' in i_str and 'EXCHANGERATES' in i_str and ',3,' in i_str:
                    i_vlookup_ok += 1

        if h_vlookup_ok == 30 and i_vlookup_ok == 30:
            print(f"PASS: Component 4 — All 30 Home Currency VLOOKUP (H) and Exchange Rate VLOOKUP (I) formulas correct")
            total_score += 0.15
        elif h_count == 30 and i_count == 30:
            print(f"PARTIAL: Component 4 — Cells filled ({h_count}/30 H, {i_count}/30 I) but VLOOKUP patterns: H={h_vlookup_ok}/30, I={i_vlookup_ok}/30")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — H column: {h_count}/30 filled, I column: {i_count}/30 filled (need VLOOKUP formulas)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Column J — Total Home Currency = G*I (0.15 points)
    # Formula should be =Grow*Irow for rows 2-31
    try:
        j_count = 0
        j_formula_ok = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=10).value  # Column J
            if val is not None:
                j_count += 1
                val_str = str(val).strip().upper().replace(' ', '')
                expected = f'=G{row}*I{row}'
                if val_str == expected.upper():
                    j_formula_ok += 1
        if j_formula_ok == 30:
            print(f"PASS: Component 5 — All 30 Total Home Currency formulas (=Grow*Irow) present and correct ({j_formula_ok}/30)")
            total_score += 0.15
        elif j_count == 30:
            print(f"PARTIAL: Component 5 — 30 cells filled but only {j_formula_ok}/30 match =Grow*Irow pattern")
            total_score += 0.08
        else:
            print(f"FAIL: Component 5 — Only {j_count}/30 Total Home Currency cells filled (need formula =Grow*Irow)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Column L — Discrepancy Flag using IF(ABS(G-K)>100,"Review","") (0.10 points)
    try:
        l_count = 0
        l_formula_ok = 0
        for row in range(2, 32):
            val = ws.cell(row=row, column=12).value  # Column L
            if val is not None:
                l_count += 1
                val_str = str(val).strip().upper().replace(' ', '')
                # Check for IF(ABS(Grow-Krow)>100 pattern
                if ('IF(' in val_str or val_str.startswith('=IF(')) and 'ABS(' in val_str and '>100' in val_str:
                    l_formula_ok += 1
        if l_formula_ok == 30:
            print(f"PASS: Component 6 — All 30 Discrepancy Flag formulas (IF/ABS) present and correct ({l_formula_ok}/30)")
            total_score += 0.10
        elif l_count == 30:
            print(f"PARTIAL: Component 6 — 30 cells filled but only {l_formula_ok}/30 contain IF/ABS/>100 pattern")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Only {l_count}/30 Discrepancy Flag cells filled (need IF/ABS formula)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Number formatting — G should be currency ($#,##0.00), J should be numeric (0.10 points)
    # In initial file G and J have 'General' format; golden applies '$#,##0.00' to G and '#,##0.00' to J
    try:
        g_fmt_ok = 0
        j_fmt_ok = 0
        # Check a sample of cells in G and J for currency format
        for row in range(2, 32):
            g_fmt = ws.cell(row=row, column=7).number_format   # Column G
            j_fmt = ws.cell(row=row, column=10).number_format  # Column J
            if g_fmt and g_fmt != 'General' and ('0.00' in g_fmt or '#' in g_fmt):
                g_fmt_ok += 1
            if j_fmt and j_fmt != 'General' and ('0.00' in j_fmt or '#' in j_fmt):
                j_fmt_ok += 1

        if g_fmt_ok >= 25 and j_fmt_ok >= 25:
            print(f"PASS: Component 7 — Currency formatting applied to G ({g_fmt_ok}/30) and J ({j_fmt_ok}/30) columns")
            total_score += 0.10
        elif g_fmt_ok >= 25:
            print(f"PARTIAL: Component 7 — Currency formatting on G ({g_fmt_ok}/30) but J only ({j_fmt_ok}/30)")
            total_score += 0.05
        elif j_fmt_ok >= 25:
            print(f"PARTIAL: Component 7 — Currency formatting on J ({j_fmt_ok}/30) but G only ({g_fmt_ok}/30)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — G formatted: {g_fmt_ok}/30, J formatted: {j_fmt_ok}/30 (need non-General number format)")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
