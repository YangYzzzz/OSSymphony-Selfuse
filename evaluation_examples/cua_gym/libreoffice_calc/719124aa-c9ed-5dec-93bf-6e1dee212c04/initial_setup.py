"""
Initial Setup: Calculate tuition and fees for international students
Task ID: calc_edu_international_student_fees_063
Domain: libreoffice_calc

Creates a spreadsheet with:
- Sheet 'IntlStudents': 30 students with Student ID, Name, Home Country,
  Domestic Tuition (D) and Billed USD (K) filled; columns E,F,G,H,I,J,L empty
- Sheet 'ExchangeRates': 15 countries with currency codes and exchange rates
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_international_student_fees_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet 1: IntlStudents
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = 'IntlStudents'

    # Headers (row 1)
    headers = [
        'Student ID',       # A
        'Name',             # B
        'Home Country',     # C
        'Domestic Tuition', # D
        "Int'l Tuition",    # E  -- EMPTY in initial
        "Int'l Fee",        # F  -- EMPTY in initial
        'Total USD',        # G  -- EMPTY in initial
        'Home Currency',    # H  -- EMPTY in initial
        'Exchange Rate',    # I  -- EMPTY in initial
        'Total Home Currency',  # J  -- EMPTY in initial
        'Billed USD',       # K
        'Discrepancy Flag', # L  -- EMPTY in initial
    ]
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Student data: (Student ID, Name, Home Country, Domestic Tuition, Billed USD)
    # 30 students from 15 different countries (2 per country)
    # Domestic tuition ranges from $8,000 to $18,000
    # Billed USD may differ slightly from calculated (some will trigger Review flag)
    students = [
        ('S2025001', 'Li Wei',          'China',        12000.00, 31500.00),
        ('S2025002', 'Zhang Mei',        'China',        14500.00, 38000.00),
        ('S2025003', 'Priya Sharma',     'India',        11000.00, 29800.00),
        ('S2025004', 'Arjun Patel',      'India',        13500.00, 35000.00),
        ('S2025005', 'Yuki Tanaka',      'Japan',        15000.00, 38850.00),
        ('S2025006', 'Kenji Nakamura',   'Japan',        10500.00, 27750.00),
        ('S2025007', 'Min-jun Lee',      'South Korea',  12500.00, 32750.00),
        ('S2025008', 'Ji-yeon Park',     'South Korea',   9500.00, 25250.00),
        ('S2025009', 'Nguyen Thi Lan',   'Vietnam',       8000.00, 21500.00),
        ('S2025010', 'Tran Van Duc',     'Vietnam',       9000.00, 24000.00),
        ('S2025011', 'Fatima Al-Hassan', 'Saudi Arabia', 16000.00, 41500.00),
        ('S2025012', 'Omar Al-Rashidi',  'Saudi Arabia', 18000.00, 46500.00),
        ('S2025013', 'Ana Garcia',       'Brazil',       11500.00, 30250.00),
        ('S2025014', 'Carlos Oliveira',  'Brazil',       13000.00, 34000.00),
        ('S2025015', 'Emre Yilmaz',      'Turkey',       10000.00, 26500.00),
        ('S2025016', 'Selin Demir',      'Turkey',       12000.00, 31700.00),  # slight discrepancy
        ('S2025017', 'Andrei Popescu',   'Romania',       8500.00, 22750.00),
        ('S2025018', 'Maria Ionescu',    'Romania',       9500.00, 25000.00),  # slight discrepancy
        ('S2025019', 'Ahmed Benali',     'Morocco',      10500.00, 27750.00),
        ('S2025020', 'Layla Mansouri',   'Morocco',      11000.00, 29000.00),
        ('S2025021', 'Ivan Petrov',      'Russia',       14000.00, 36500.00),
        ('S2025022', 'Olga Sokolova',    'Russia',       13000.00, 34000.00),  # discrepancy >100
        ('S2025023', 'Sophia Müller',    'Germany',      15500.00, 40250.00),  # discrepancy >100
        ('S2025024', 'Felix Bauer',      'Germany',      17000.00, 44000.00),
        ('S2025025', 'Emma Dubois',      'France',       16500.00, 43000.00),  # discrepancy >100
        ('S2025026', 'Antoine Leroy',    'France',       14000.00, 36500.00),
        ('S2025027', 'Aisha Okonkwo',    'Nigeria',       9000.00, 24000.00),
        ('S2025028', 'Chidi Eze',        'Nigeria',       8000.00, 21500.00),
        ('S2025029', 'Valentina Cruz',   'Mexico',       10000.00, 26500.00),
        ('S2025030', 'Miguel Santos',    'Mexico',       11500.00, 30500.00),
    ]

    for row_num, (sid, name, country, dom_tuition, billed) in enumerate(students, 2):
        ws1.cell(row=row_num, column=1, value=sid)
        ws1.cell(row=row_num, column=2, value=name)
        ws1.cell(row=row_num, column=3, value=country)
        ws1.cell(row=row_num, column=4, value=dom_tuition)
        # Columns E(5), F(6), G(7), H(8), I(9), J(10) are LEFT EMPTY
        ws1.cell(row=row_num, column=11, value=billed)
        # Column L(12) is LEFT EMPTY

    # Format column D and K as currency in initial (they are filled)
    currency_fmt = '$#,##0.00'
    for row_num in range(2, 32):
        ws1.cell(row=row_num, column=4).number_format = currency_fmt
        ws1.cell(row=row_num, column=11).number_format = currency_fmt

    # Column widths
    col_widths = {
        'A': 12, 'B': 22, 'C': 16, 'D': 18,
        'E': 15, 'F': 12, 'G': 14, 'H': 16,
        'I': 14, 'J': 20, 'K': 14, 'L': 18,
    }
    for col_letter, width in col_widths.items():
        ws1.column_dimensions[col_letter].width = width

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # ------------------------------------------------------------------ #
    # Sheet 2: ExchangeRates
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('ExchangeRates')

    # Headers
    er_headers = ['Country', 'Currency Code', 'Rate (per USD)']
    er_header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    er_header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    for col, h in enumerate(er_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = er_header_font
        cell.fill = er_header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 15 countries with realistic exchange rates (as of early 2025)
    exchange_rates = [
        ('Brazil',       'BRL',  5.05),
        ('China',        'CNY',  7.24),
        ('France',       'EUR',  0.92),
        ('Germany',      'EUR',  0.92),
        ('India',        'INR', 83.12),
        ('Japan',        'JPY', 149.85),
        ('Mexico',       'MXN', 17.15),
        ('Morocco',      'MAD',  9.98),
        ('Nigeria',      'NGN', 1550.00),
        ('Romania',      'RON',  4.65),
        ('Russia',       'RUB', 88.50),
        ('Saudi Arabia', 'SAR',  3.75),
        ('South Korea',  'KRW', 1325.40),
        ('Turkey',       'TRY', 32.10),
        ('Vietnam',      'VND', 24850.00),
    ]

    for row_num, (country, code, rate) in enumerate(exchange_rates, 2):
        ws2.cell(row=row_num, column=1, value=country)
        ws2.cell(row=row_num, column=2, value=code)
        ws2.cell(row=row_num, column=3, value=rate)

    # Column widths
    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  - Sheet IntlStudents: 30 students, columns E,F,G,H,I,J,L empty')
    print(f'  - Sheet ExchangeRates: 15 countries with currency codes and rates')


create_initial()
