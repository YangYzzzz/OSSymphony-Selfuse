"""
Initial Setup: Customer LTV Calculation - AccountData with 75 accounts
Task ID: calc_sales_customer_ltv_022
Domain: libreoffice_calc

Creates an initial spreadsheet with AccountData sheet containing:
- 75 accounts with Account ID, Company, Avg Monthly Spend, Retention Months
- LTV (E) and LTV Rank (F) columns are EMPTY (task requires calculating these)
- No formulas, no conditional formatting, no sorting by LTV
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_customer_ltv_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: AccountData ---
    ws = wb.active
    ws.title = 'AccountData'

    # Headers in Row 1
    headers = ['Account ID', 'Company', 'Avg Monthly Spend', 'Retention Months', 'LTV', 'LTV Rank']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 75 accounts with realistic company names and data
    # Monthly spend: $5,000 - $85,000; Retention: 6-60 months
    accounts = [
        ('ACC-001', 'Meridian Technologies Inc.', 72000, 48),
        ('ACC-002', 'BlueCrest Financial Group', 45000, 36),
        ('ACC-003', 'Apex Solutions Corp', 83000, 54),
        ('ACC-004', 'Northgate Enterprises', 31000, 24),
        ('ACC-005', 'Silverline Consulting LLC', 58000, 42),
        ('ACC-006', 'Orion Systems Ltd', 27000, 18),
        ('ACC-007', 'Cascade Digital Partners', 65000, 60),
        ('ACC-008', 'Redwood Analytics', 19000, 12),
        ('ACC-009', 'Summit Healthcare Solutions', 77000, 52),
        ('ACC-010', 'Lakeside Manufacturing Co', 34000, 30),
        ('ACC-011', 'Ironclad Security Services', 51000, 38),
        ('ACC-012', 'Greenfield Biotech', 85000, 56),
        ('ACC-013', 'Vanguard Logistics', 22000, 20),
        ('ACC-014', 'Pinnacle Retail Group', 48000, 44),
        ('ACC-015', 'Horizon Energy Partners', 69000, 50),
        ('ACC-016', 'Coastal Properties LLC', 15000, 10),
        ('ACC-017', 'TechEdge Innovations', 80000, 58),
        ('ACC-018', 'Brightwater Media', 38000, 28),
        ('ACC-019', 'Pacific Ventures Capital', 55000, 46),
        ('ACC-020', 'Quantum Computing Solutions', 74000, 55),
        ('ACC-021', 'Maple Grove Industries', 28000, 22),
        ('ACC-022', 'Starfield Communications', 63000, 49),
        ('ACC-023', 'Aurora Education Group', 17000, 14),
        ('ACC-024', 'NexGen Pharmaceuticals', 82000, 57),
        ('ACC-025', 'Clearwater Environmental', 40000, 32),
        ('ACC-026', 'Titanium Auto Parts', 53000, 40),
        ('ACC-027', 'Eastern Seaboard Trust', 71000, 51),
        ('ACC-028', 'CloudNine Software Ltd', 33000, 26),
        ('ACC-029', 'Frontier Agriculture Co', 24000, 16),
        ('ACC-030', 'MetroLink Transportation', 60000, 47),
        ('ACC-031', 'Sterling Asset Management', 78000, 53),
        ('ACC-032', 'Wildfire Marketing Agency', 42000, 34),
        ('ACC-033', 'Unified Health Systems', 67000, 48),
        ('ACC-034', 'PrimeSource Staffing', 20000, 15),
        ('ACC-035', 'Galactic Entertainment', 56000, 43),
        ('ACC-036', 'Ironwood Construction', 36000, 29),
        ('ACC-037', 'SkyBridge Telecom', 73000, 54),
        ('ACC-038', 'Fulcrum Dynamics', 47000, 37),
        ('ACC-039', 'Emberstone Hospitality', 62000, 46),
        ('ACC-040', 'Nexus Investment Partners', 30000, 23),
        ('ACC-041', 'Radiant Solar Energy', 84000, 59),
        ('ACC-042', 'BluePeak Advisors', 25000, 19),
        ('ACC-043', 'Continuum Insurance Group', 68000, 50),
        ('ACC-044', 'Ashwood Publishing House', 14000, 9),
        ('ACC-045', 'Fortress Data Centers', 76000, 55),
        ('ACC-046', 'Meadowbrook Retail', 35000, 27),
        ('ACC-047', 'Cobalt Cybersecurity', 57000, 44),
        ('ACC-048', 'Highpoint Legal Partners', 43000, 35),
        ('ACC-049', 'Cascade River Logistics', 70000, 52),
        ('ACC-050', 'Whiterock Minerals LLC', 18000, 13),
        ('ACC-051', 'Catalyst Growth Fund', 81000, 57),
        ('ACC-052', 'DawnBreaker Technology', 29000, 21),
        ('ACC-053', 'Ember Flyte Aviation', 64000, 48),
        ('ACC-054', 'Springboard Ventures', 37000, 31),
        ('ACC-055', 'Perimeter Defense Systems', 75000, 56),
        ('ACC-056', 'Coral Reef Travel Agency', 23000, 17),
        ('ACC-057', 'Momentum Capital Group', 79000, 58),
        ('ACC-058', 'Horizon Ridge Properties', 46000, 38),
        ('ACC-059', 'Prism Analytics Co', 61000, 45),
        ('ACC-060', 'Keystone Building Services', 32000, 25),
        ('ACC-061', 'Northern Star Renewables', 85000, 60),
        ('ACC-062', 'Verdant Farms Inc', 16000, 11),
        ('ACC-063', 'Colosseum Events LLC', 54000, 41),
        ('ACC-064', 'Arclight Communications', 69000, 53),
        ('ACC-065', 'Bridgeport Finance Corp', 41000, 33),
        ('ACC-066', 'TerraVerde Landscaping', 26000, 20),
        ('ACC-067', 'Pathfinder AI Solutions', 77000, 56),
        ('ACC-068', 'Goldfield Mining Group', 49000, 39),
        ('ACC-069', 'Coastal Star Media', 66000, 49),
        ('ACC-070', 'Zenith Pharmaceuticals', 83000, 59),
        ('ACC-071', 'Pinnacle Sports Agency', 21000, 16),
        ('ACC-072', 'Westside Capital Partners', 72000, 54),
        ('ACC-073', 'Granite Peak Software', 39000, 32),
        ('ACC-074', 'Luminary Design Studio', 59000, 45),
        ('ACC-075', 'Streamline Operations Co', 44000, 35),
    ]

    for r, (acc_id, company, monthly_spend, retention) in enumerate(accounts, 2):
        ws.cell(row=r, column=1, value=acc_id)
        ws.cell(row=r, column=2, value=company)
        ws.cell(row=r, column=3, value=monthly_spend)
        ws.cell(row=r, column=4, value=retention)
        # Columns E (LTV) and F (LTV Rank) intentionally left empty

    # Apply currency format to column C (Avg Monthly Spend)
    for r in range(2, 77):
        ws.cell(row=r, column=3).number_format = '$#,##0'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: AccountData')
    print(f'  Rows: 1 header + 75 data rows')
    print(f'  Columns A-D populated, E-F empty (as required)')


create_initial()
