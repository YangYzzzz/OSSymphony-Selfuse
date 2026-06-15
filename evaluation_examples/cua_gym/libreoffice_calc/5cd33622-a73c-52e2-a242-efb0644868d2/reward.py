"""
Reward Script: Calculate customer lifetime value, RANK, conditional formatting, sort
Task ID: calc_sales_customer_ltv_022
Domain: libreoffice_calc
Scoring:
  Component 1: LTV formulas in E2:E76 (=Cn*Dn pattern)           0.30
  Component 2: RANK formulas in F2:F76 (RANK over $E$2:$E$76)    0.20
  Component 3: E column number format ($#,##0 currency)           0.10
  Component 4: Conditional formatting (gold fill, F<=10, A2:F76)  0.25
  Component 5: Data sorted descending by LTV (C*D)                0.15
  Total:                                                           1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_sales_customer_ltv_022'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Validate sheet exists
    if 'AccountData' not in wb.sheetnames:
        print("FAIL: Sheet 'AccountData' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['AccountData']

    # -----------------------------------------------------------------------
    # Component 1: LTV formulas in E2:E76 (=Cn*Dn pattern) — 0.30 points
    # Task requires: E2:E76 should contain =C2*D2 style formulas
    # This FAILS on initial (E column is empty) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        ltv_formula_count = 0
        ltv_formula_correct = 0

        for row in range(2, 77):
            e_val = ws.cell(row=row, column=5).value
            if e_val is not None:
                ltv_formula_count += 1
                # Accept =Cn*Dn pattern (e.g. =C2*D2, =C10*D10)
                if isinstance(e_val, str):
                    normalized = e_val.strip().upper().replace(' ', '')
                    expected = f'=C{row}*D{row}'
                    # Also accept reverse order D*C
                    alt_expected = f'=D{row}*C{row}'
                    if normalized == expected.upper() or normalized == alt_expected.upper():
                        ltv_formula_correct += 1

        if ltv_formula_count == 75 and ltv_formula_correct == 75:
            print(f"PASS: Component 1 — All 75 LTV formulas (=Cn*Dn) found in E2:E76 (0.30 pts)")
            total_score += 0.30
        elif ltv_formula_count == 75 and ltv_formula_correct >= 70:
            # Most formulas correct — partial credit
            print(f"PASS (partial): Component 1 — {ltv_formula_correct}/75 LTV formulas correct (0.20 pts)")
            total_score += 0.20
        elif ltv_formula_count > 0:
            print(f"FAIL: Component 1 — Only {ltv_formula_count}/75 cells populated in E column, "
                  f"{ltv_formula_correct} with correct formula pattern")
        else:
            print("FAIL: Component 1 — E column (LTV) is empty, no formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: RANK formulas in F2:F76 — 0.20 points
    # Task requires: F2:F76 should contain =RANK(En,$E$2:$E$76,0)
    # This FAILS on initial (F column is empty) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        rank_formula_count = 0
        rank_formula_correct = 0

        for row in range(2, 77):
            f_val = ws.cell(row=row, column=6).value
            if f_val is not None:
                rank_formula_count += 1
                if isinstance(f_val, str):
                    normalized = f_val.strip().upper().replace(' ', '')
                    # Expected pattern: =RANK(En,$E$2:$E$76,0)
                    expected = f'=RANK(E{row},$E$2:$E$76,0)'
                    if normalized == expected.upper():
                        rank_formula_correct += 1

        if rank_formula_count == 75 and rank_formula_correct == 75:
            print(f"PASS: Component 2 — All 75 RANK formulas found in F2:F76 (0.20 pts)")
            total_score += 0.20
        elif rank_formula_count == 75 and rank_formula_correct >= 70:
            print(f"PASS (partial): Component 2 — {rank_formula_correct}/75 RANK formulas correct (0.12 pts)")
            total_score += 0.12
        elif rank_formula_count > 0:
            print(f"FAIL: Component 2 — Only {rank_formula_count}/75 cells populated in F column, "
                  f"{rank_formula_correct} with correct RANK formula pattern")
        else:
            print("FAIL: Component 2 — F column (LTV Rank) is empty, no formulas found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: E column number format ($#,##0 currency) — 0.10 points
    # Task requires: LTV column formatted as currency $#,##0
    # This FAILS on initial (E column is empty/General) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        currency_format_count = 0
        # Check a sample of E cells (rows 2, 10, 20, 40, 76)
        sample_rows = [2, 10, 20, 40, 76]
        for row in sample_rows:
            nf = ws.cell(row=row, column=5).number_format
            # Accept $#,##0 or similar currency formats
            if nf and ('$' in nf or 'USD' in nf.upper()):
                currency_format_count += 1

        # Also check all rows for thoroughness
        all_currency_count = 0
        for row in range(2, 77):
            nf = ws.cell(row=row, column=5).number_format
            if nf and '$' in nf:
                all_currency_count += 1

        if all_currency_count == 75:
            print(f"PASS: Component 3 — All 75 E-column cells have currency format ($#,##0) (0.10 pts)")
            total_score += 0.10
        elif all_currency_count >= 70:
            print(f"PASS (partial): Component 3 — {all_currency_count}/75 E-column cells have currency format (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — Only {all_currency_count}/75 E-column cells have currency format "
                  f"(expected: $#,##0 for all, E2 has: {ws.cell(row=2, column=5).number_format!r})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional formatting (gold fill, F<=10, range A2:F76) — 0.25 points
    # Task requires: highlight top 10 (rank <= 10) with gold/yellow background on entire row
    # This FAILS on initial (no CF rules) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        cf_found = False
        cf_correct_range = False
        cf_correct_formula = False
        cf_correct_color = False

        cf_rules_dict = ws.conditional_formatting._cf_rules
        if len(cf_rules_dict) > 0:
            cf_found = True

            for cf_range_obj, rules_list in cf_rules_dict.items():
                range_str = str(cf_range_obj)
                # Check if range covers rows 2-76 (A2:F76 or similar)
                if 'A2' in range_str and '76' in range_str:
                    cf_correct_range = True

                for rule in rules_list:
                    # Check formula references F column <= 10
                    if hasattr(rule, 'formula') and rule.formula:
                        for formula in rule.formula:
                            formula_upper = str(formula).upper().replace(' ', '')
                            # Accept $F2<=10 or F2<=10 or similar
                            if 'F' in formula_upper and '10' in formula_upper and '<=' in formula_upper:
                                cf_correct_formula = True

                    # Check fill color is gold/yellow
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fg_color = rule.dxf.fill.fgColor.rgb
                            if fg_color:
                                # Gold/yellow colors: FFD700 (gold), FFFF00 (yellow), FFD966 (Excel yellow)
                                # Accept any gold/yellow-ish color
                                color_upper = fg_color.upper()
                                # FFFFD700 = gold, FFFFFF00 = yellow, FFFFD966 = Excel gold, FFFFC000 = orange-gold
                                gold_colors = ['FFFFD700', 'FFFFFF00', 'FFFFD966', 'FFFFC000',
                                               'FFFFCC00', 'FFFFF200', 'FFFE9A00']
                                if color_upper in gold_colors:
                                    cf_correct_color = True
                                elif color_upper.endswith('FFD700') or color_upper.endswith('FFD966'):
                                    cf_correct_color = True
                                # Accept anything that looks gold/yellowish
                                elif len(fg_color) >= 6:
                                    # Parse RGB: high R, high G, low B = yellow/gold
                                    rgb_part = color_upper[-6:]
                                    r_val = int(rgb_part[0:2], 16)
                                    g_val = int(rgb_part[2:4], 16)
                                    b_val = int(rgb_part[4:6], 16)
                                    if r_val >= 180 and g_val >= 160 and b_val <= 100:
                                        cf_correct_color = True
                        except Exception as color_err:
                            print(f"  Note: could not parse CF fill color: {color_err}")

        if cf_found and cf_correct_range and cf_correct_formula and cf_correct_color:
            print(f"PASS: Component 4 — Conditional formatting found: range A2:F76, "
                  f"formula F<=10, gold/yellow fill (0.25 pts)")
            total_score += 0.25
        elif cf_found and cf_correct_formula and cf_correct_color:
            print(f"PASS (partial): Component 4 — CF found with correct formula and color "
                  f"but range check failed (0.15 pts). cf_correct_range={cf_correct_range}")
            total_score += 0.15
        elif cf_found and (cf_correct_formula or cf_correct_color):
            components = []
            if cf_correct_formula:
                components.append("correct formula")
            if cf_correct_color:
                components.append("correct color")
            print(f"PASS (partial): Component 4 — CF found with {', '.join(components)} (0.10 pts)")
            total_score += 0.10
        elif cf_found:
            print(f"FAIL: Component 4 — CF rule found but incorrect: "
                  f"range_ok={cf_correct_range}, formula_ok={cf_correct_formula}, color_ok={cf_correct_color}")
        else:
            print("FAIL: Component 4 — No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Data sorted descending by LTV (C*D) — 0.15 points
    # Task requires: sort data by LTV (column E) descending
    # This FAILS on initial (data is in original unsorted order) and PASSES on golden
    # -----------------------------------------------------------------------
    try:
        # Compute LTV (C*D) for each data row and check descending order
        ltvs = []
        valid_rows = 0

        for row in range(2, 77):
            c_val = ws.cell(row=row, column=3).value
            d_val = ws.cell(row=row, column=4).value
            if c_val is not None and d_val is not None:
                try:
                    ltv = float(c_val) * float(d_val)
                    ltvs.append(ltv)
                    valid_rows += 1
                except (ValueError, TypeError):
                    ltvs.append(None)

        if valid_rows < 70:
            print(f"FAIL: Component 5 — Only {valid_rows} rows with valid C*D values (expected 75)")
        else:
            # Check descending sort (filter out None values for comparison)
            valid_ltvs = [v for v in ltvs if v is not None]
            is_sorted_desc = all(
                valid_ltvs[i] >= valid_ltvs[i + 1]
                for i in range(len(valid_ltvs) - 1)
            )

            if is_sorted_desc:
                print(f"PASS: Component 5 — Data is sorted descending by LTV (C*D). "
                      f"Top LTV: {valid_ltvs[0]:,.0f}, Bottom LTV: {valid_ltvs[-1]:,.0f} (0.15 pts)")
                total_score += 0.15
            else:
                # Count inversions to see how well sorted it is
                inversions = sum(
                    1 for i in range(len(valid_ltvs) - 1)
                    if valid_ltvs[i] < valid_ltvs[i + 1]
                )
                print(f"FAIL: Component 5 — Data is NOT sorted descending by LTV. "
                      f"{inversions} out-of-order pairs found (expected 0)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
