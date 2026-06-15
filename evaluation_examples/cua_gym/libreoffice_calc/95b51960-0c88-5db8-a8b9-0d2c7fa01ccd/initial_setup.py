"""
Initial Setup: Save workbook as ODS file to Documents folder
Task ID: calc_gg1_040
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gg1_040'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Style definitions ---
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'
    date_fmt = 'yyyy-mm-dd'
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # === Sheet 1: Budget Overview (monthly budget data) ===
    ws1 = wb.active
    ws1.title = "Budget Overview"

    headers1 = ["Month", "Revenue", "COGS", "Gross Profit", "Operating Expenses",
                 "Marketing", "Salaries", "Rent", "Utilities", "Net Income"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    months = ["Jan 2025", "Feb 2025", "Mar 2025", "Apr 2025", "May 2025", "Jun 2025",
              "Jul 2025", "Aug 2025", "Sep 2025", "Oct 2025", "Nov 2025", "Dec 2025"]

    revenue_data = [245000, 263500, 278900, 291200, 305800, 312400,
                    298700, 319500, 334200, 348900, 367100, 425600]
    cogs_data = [98000, 105400, 111560, 116480, 122320, 124960,
                 119480, 127800, 133680, 139560, 146840, 170240]
    opex_data = [32000, 33500, 34000, 31500, 35200, 36800,
                 33000, 34500, 35800, 37200, 38500, 42000]
    marketing_data = [18500, 22000, 25000, 19800, 27500, 30000,
                      21000, 24500, 28000, 31500, 35000, 45000]
    salaries_data = [85000, 85000, 85000, 87000, 87000, 87000,
                     89000, 89000, 89000, 91000, 91000, 91000]
    rent = 12500
    utilities_data = [3200, 3100, 3400, 3600, 4100, 4500,
                      4800, 4600, 4200, 3800, 3500, 3300]

    for i, month in enumerate(months):
        r = i + 2
        ws1.cell(row=r, column=1, value=month).border = thin_border
        ws1.cell(row=r, column=2, value=revenue_data[i]).number_format = currency_fmt
        ws1.cell(row=r, column=2).border = thin_border
        ws1.cell(row=r, column=3, value=cogs_data[i]).number_format = currency_fmt
        ws1.cell(row=r, column=3).border = thin_border
        # Gross Profit = Revenue - COGS (formula)
        ws1.cell(row=r, column=4, value=f'=B{r}-C{r}').number_format = currency_fmt
        ws1.cell(row=r, column=4).border = thin_border
        ws1.cell(row=r, column=5, value=opex_data[i]).number_format = currency_fmt
        ws1.cell(row=r, column=5).border = thin_border
        ws1.cell(row=r, column=6, value=marketing_data[i]).number_format = currency_fmt
        ws1.cell(row=r, column=6).border = thin_border
        ws1.cell(row=r, column=7, value=salaries_data[i]).number_format = currency_fmt
        ws1.cell(row=r, column=7).border = thin_border
        ws1.cell(row=r, column=8, value=rent).number_format = currency_fmt
        ws1.cell(row=r, column=8).border = thin_border
        ws1.cell(row=r, column=9, value=utilities_data[i]).number_format = currency_fmt
        ws1.cell(row=r, column=9).border = thin_border
        # Net Income = Gross Profit - OpEx - Marketing - Salaries - Rent - Utilities
        ws1.cell(row=r, column=10, value=f'=D{r}-E{r}-F{r}-G{r}-H{r}-I{r}').number_format = currency_fmt
        ws1.cell(row=r, column=10).border = thin_border

    # Totals row
    total_row = 14
    ws1.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws1.cell(row=total_row, column=1).border = thin_border
    for col in range(2, 11):
        from openpyxl.utils import get_column_letter
        cl = get_column_letter(col)
        ws1.cell(row=total_row, column=col, value=f'=SUM({cl}2:{cl}13)').number_format = currency_fmt
        ws1.cell(row=total_row, column=col).font = Font(bold=True)
        ws1.cell(row=total_row, column=col).border = thin_border

    # Set column widths
    ws1.column_dimensions['A'].width = 14
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws1.column_dimensions[col_letter].width = 16
    ws1.freeze_panes = "A2"

    # === Sheet 2: Department Breakdown ===
    ws2 = wb.create_sheet("Department Breakdown")

    headers2 = ["Department", "Q1 Budget", "Q2 Budget", "Q3 Budget", "Q4 Budget",
                 "Annual Total", "Headcount", "Cost Per Head"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    departments = [
        ["Engineering", 125000, 132000, 138500, 145000, None, 18, None],
        ["Marketing", 65500, 72300, 68500, 85500, None, 8, None],
        ["Sales", 95000, 98500, 102000, 115000, None, 12, None],
        ["Human Resources", 42000, 43500, 44000, 46000, None, 5, None],
        ["Finance", 38000, 39200, 40500, 42800, None, 6, None],
        ["Operations", 78000, 81500, 84000, 89500, None, 10, None],
        ["Customer Support", 55000, 58000, 61500, 67000, None, 9, None],
        ["Legal", 35000, 36500, 37800, 39200, None, 4, None],
        ["IT Infrastructure", 88000, 92000, 95500, 101000, None, 7, None],
        ["Executive", 110000, 110000, 115000, 120000, None, 3, None],
    ]

    for i, dept_row in enumerate(departments):
        r = i + 2
        ws2.cell(row=r, column=1, value=dept_row[0]).border = thin_border
        for col_idx in range(1, 5):
            ws2.cell(row=r, column=col_idx + 1, value=dept_row[col_idx]).number_format = currency_fmt
            ws2.cell(row=r, column=col_idx + 1).border = thin_border
        # Annual Total = sum of quarterly budgets
        ws2.cell(row=r, column=6, value=f'=SUM(B{r}:E{r})').number_format = currency_fmt
        ws2.cell(row=r, column=6).border = thin_border
        ws2.cell(row=r, column=7, value=dept_row[6]).border = thin_border
        # Cost Per Head = Annual Total / Headcount
        ws2.cell(row=r, column=8, value=f'=F{r}/G{r}').number_format = currency_fmt
        ws2.cell(row=r, column=8).border = thin_border

    # Department totals
    dept_total_row = len(departments) + 2
    ws2.cell(row=dept_total_row, column=1, value="TOTAL").font = Font(bold=True)
    ws2.cell(row=dept_total_row, column=1).border = thin_border
    for col in range(2, 9):
        cl = get_column_letter(col)
        ws2.cell(row=dept_total_row, column=col,
                 value=f'=SUM({cl}2:{cl}{dept_total_row-1})').number_format = currency_fmt
        ws2.cell(row=dept_total_row, column=col).font = Font(bold=True)
        ws2.cell(row=dept_total_row, column=col).border = thin_border

    ws2.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws2.column_dimensions[col_letter].width = 16
    ws2.freeze_panes = "A2"

    # === Sheet 3: Quarterly Summary ===
    ws3 = wb.create_sheet("Quarterly Summary")

    headers3 = ["Quarter", "Total Revenue", "Total Expenses", "Net Profit",
                 "Profit Margin", "YoY Growth"]
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    quarters = [
        ["Q1 2025", 787400, 612760, None, None, 0.082],
        ["Q2 2025", 909400, 703240, None, None, 0.115],
        ["Q3 2025", 952400, 731460, None, None, 0.094],
        ["Q4 2025", 1141600, 898140, None, None, 0.138],
    ]

    for i, q in enumerate(quarters):
        r = i + 2
        ws3.cell(row=r, column=1, value=q[0]).border = thin_border
        ws3.cell(row=r, column=2, value=q[1]).number_format = currency_fmt
        ws3.cell(row=r, column=2).border = thin_border
        ws3.cell(row=r, column=3, value=q[2]).number_format = currency_fmt
        ws3.cell(row=r, column=3).border = thin_border
        # Net Profit = Revenue - Expenses
        ws3.cell(row=r, column=4, value=f'=B{r}-C{r}').number_format = currency_fmt
        ws3.cell(row=r, column=4).border = thin_border
        # Profit Margin = Net Profit / Revenue
        ws3.cell(row=r, column=5, value=f'=D{r}/B{r}').number_format = pct_fmt
        ws3.cell(row=r, column=5).border = thin_border
        ws3.cell(row=r, column=6, value=q[5]).number_format = pct_fmt
        ws3.cell(row=r, column=6).border = thin_border

    # Annual summary row
    annual_row = 6
    ws3.cell(row=annual_row, column=1, value="FY 2025 Total").font = Font(bold=True)
    ws3.cell(row=annual_row, column=1).border = thin_border
    ws3.cell(row=annual_row, column=2, value='=SUM(B2:B5)').number_format = currency_fmt
    ws3.cell(row=annual_row, column=2).font = Font(bold=True)
    ws3.cell(row=annual_row, column=2).border = thin_border
    ws3.cell(row=annual_row, column=3, value='=SUM(C2:C5)').number_format = currency_fmt
    ws3.cell(row=annual_row, column=3).font = Font(bold=True)
    ws3.cell(row=annual_row, column=3).border = thin_border
    ws3.cell(row=annual_row, column=4, value='=B6-C6').number_format = currency_fmt
    ws3.cell(row=annual_row, column=4).font = Font(bold=True)
    ws3.cell(row=annual_row, column=4).border = thin_border
    ws3.cell(row=annual_row, column=5, value='=D6/B6').number_format = pct_fmt
    ws3.cell(row=annual_row, column=5).font = Font(bold=True)
    ws3.cell(row=annual_row, column=5).border = thin_border

    ws3.column_dimensions['A'].width = 16
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws3.column_dimensions[col_letter].width = 18
    ws3.freeze_panes = "A2"

    # Ensure /home/user/Documents/ exists but does NOT contain budget_final.ods
    os.makedirs('/home/user/Documents', exist_ok=True)
    ods_path = '/home/user/Documents/budget_final.ods'
    if os.path.exists(ods_path):
        os.remove(ods_path)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
