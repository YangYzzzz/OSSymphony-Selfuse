"""
Initial Setup: Payroll spreadsheet for nested ROUND+SUM formula task
Task ID: calc_fmb_nested_round_sum_053
Domain: libreoffice_calc

Creates a spreadsheet with 50 employee records.
The sum of ROUND(salary, -3) for rows 2-51 = 4,217,000.
Cell E2 is EMPTY (target for =SUMPRODUCT(ROUND(C2:C51,-3))).
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_nested_round_sum_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Payroll'

    # --- Headers ---
    ws['A1'] = 'Emp ID'
    ws['B1'] = 'Name'
    ws['C1'] = 'Exact Salary'
    ws['D1'] = 'Department'

    # --- Employee data ---
    # Salaries are carefully chosen so that ROUND(salary, -3) sums to exactly 4,217,000.
    # ROUND(x, -3) rounds to nearest 1000.
    # Each salary was verified: round(salary / 1000) * 1000 gives the intended value.
    employees = [
        ('EMP001', 'Sarah Chen',           74950,  'Engineering'),
        ('EMP002', 'Marcus Johnson',       98780,  'Marketing'),
        ('EMP003', 'Priya Patel',          59420,  'Finance'),
        ('EMP004', 'Derek Okafor',        135170,  'Engineering'),
        ('EMP005', 'Lisa Nguyen',          67840,  'HR'),
        ('EMP006', 'James Whitfield',      86010,  'Engineering'),
        ('EMP007', 'Ana Gutierrez',        52680,  'Operations'),
        ('EMP008', 'Robert Kim',          101260,  'Engineering'),
        ('EMP009', 'Fatima Al-Hassan',     76930,  'Finance'),
        ('EMP010', 'Thomas Erikson',       63350,  'Marketing'),
        ('EMP011', 'Yuki Tanaka',          88790,  'Engineering'),
        ('EMP012', 'Carla Reyes',          57140,  'HR'),
        ('EMP013', 'Nathan Brooks',       123980,  'Engineering'),
        ('EMP014', 'Adaeze Okonkwo',       71620,  'Operations'),
        ('EMP015', 'Michael Torres',       85150,  'Finance'),
        ('EMP016', 'Ingrid Svensson',      95870,  'Engineering'),
        ('EMP017', 'Daniel Park',          66330,  'Marketing'),
        ('EMP018', 'Mei-Ling Zhou',        80450,  'Engineering'),
        ('EMP019', 'Oliver Bauer',         60980,  'Operations'),
        ('EMP020', 'Nadia Petrov',        118730,  'Finance'),
        ('EMP021', 'Kwame Asante',         74060,  'HR'),
        ('EMP022', 'Helena Kowalski',      87640,  'Engineering'),
        ('EMP023', 'Jordan Malik',         55470,  'Marketing'),
        ('EMP024', 'Siobhan Murphy',      109370,  'Engineering'),
        ('EMP025', 'Victor Santos',        70210,  'Operations'),
        ('EMP026', 'Aisha Diallo',         82880,  'Finance'),
        ('EMP027', 'Emmanuel Dubois',      62040,  'Marketing'),
        ('EMP028', 'Hana Yoshida',         95370,  'Engineering'),
        ('EMP029', "Patrick O'Brien",      76790,  'HR'),
        ('EMP030', 'Riya Sharma',         131150,  'Engineering'),
        ('EMP031', 'Calvin Washington',    67680,  'Operations'),
        ('EMP032', 'Aleksandra Novak',     90930,  'Finance'),
        ('EMP033', 'Miguel Castillo',      59260,  'Marketing'),
        ('EMP034', 'Diana Popescu',       106390,  'Engineering'),
        ('EMP035', 'Bashir Ahmed',         83250,  'HR'),
        ('EMP036', 'Svetlana Ivanova',     84470,  'Engineering'),
        ('EMP037', 'Leon Fischer',         63940,  'Operations'),
        ('EMP038', 'Amara Diop',           98080,  'Finance'),
        ('EMP039', 'Wesley Coleman',       84950,  'Marketing'),
        ('EMP040', 'Yara El-Amin',         89150,  'Engineering'),
        ('EMP041', 'Finn Andersen',        57780,  'Operations'),
        ('EMP042', 'Catalina Ruiz',       102930,  'Engineering'),
        ('EMP043', 'Damien Laurent',       71370,  'HR'),
        ('EMP044', 'Zhen Liu',             91740,  'Finance'),
        ('EMP045', 'Imogen Clarke',        65250,  'Marketing'),
        ('EMP046', 'Rashid Omar',          92810,  'Engineering'),
        ('EMP047', 'Valentina Greco',      79180,  'Operations'),
        ('EMP048', 'Samuel Osei',         127420,  'Engineering'),
        ('EMP049', 'Brigitte Moreau',      86950,  'Finance'),
        ('EMP050', 'Hiroshi Yamamoto',    106260,  'Engineering'),
    ]

    # Verify the rounded sum equals 4,217,000
    rounded_sum = sum(round(sal / 1000) * 1000 for _, _, sal, _ in employees)
    assert rounded_sum == 4217000, f'ERROR: rounded sum is {rounded_sum}, expected 4217000'
    print(f'Verified: ROUND(salary,-3) sum = {rounded_sum:,}')

    for row_idx, (emp_id, name, salary, dept) in enumerate(employees, 2):
        ws.cell(row=row_idx, column=1, value=emp_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=salary)
        ws.cell(row=row_idx, column=4, value=dept)

    # D2 label per task context
    ws['D2'] = 'Rounded Total'

    # E2 is intentionally left EMPTY (target cell for the task)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

create_initial()
