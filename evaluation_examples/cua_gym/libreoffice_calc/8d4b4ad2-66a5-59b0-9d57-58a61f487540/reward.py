"""
Reward Script: Calculate total payroll cost with ROUND-then-SUM in E2
Task ID: calc_fmb_nested_round_sum_053
Domain: libreoffice_calc

Task: Put =SUMPRODUCT(ROUND(C2:C51,-3)) (or equivalent) in cell E2.
      The formula should round each salary to the nearest $1,000 then sum.

Scoring Rubric:
  Component 1: E2 contains a formula (not empty) — 0.3 pts
  Component 2: Formula uses ROUND with -3 and covers range C2:C51 — 0.3 pts
  Component 3: Formula uses SUMPRODUCT aggregation and the computed
               ROUND-then-SUM result equals 4,217,000 — 0.4 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_nested_round_sum_053'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook (formula mode to read formula strings)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the Payroll sheet exists (precondition gate)
    if 'Payroll' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Payroll' not found in workbook.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Payroll']

    # Component 1: E2 contains a formula (0.3 points)
    # In the initial file, E2 is None (empty).
    # After task completion, E2 must hold a formula string starting with '='.
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str) and e2_value.strip().startswith('='):
            print(f"PASS: Component 1 — E2 contains a formula: {repr(e2_value)} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — E2 should contain a formula, found: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not read E2: {e}")

    # Component 2: Formula uses ROUND with -3 and covers the full range C2:C51 (0.3 points)
    # The task requires rounding each salary to the nearest $1,000 (ROUND(..., -3))
    # covering ALL 50 employees (rows 2-51, column C).
    try:
        e2_value = ws['E2'].value
        if e2_value is not None and isinstance(e2_value, str):
            formula_upper = e2_value.upper().replace(' ', '')
            has_round = 'ROUND(' in formula_upper
            has_neg3 = '-3)' in formula_upper or ',-3)' in formula_upper
            has_range = 'C2:C51' in formula_upper
            if has_round and has_neg3 and has_range:
                print(f"PASS: Component 2 — Formula uses ROUND(...,-3) over C2:C51 (0.3 pts)")
                total_score += 0.3
            else:
                missing = []
                if not has_round:
                    missing.append('ROUND function')
                if not has_neg3:
                    missing.append('-3 (nearest 1000)')
                if not has_range:
                    missing.append('range C2:C51')
                print(f"FAIL: Component 2 — Formula missing: {', '.join(missing)}. Found: {repr(e2_value)}")
        else:
            print(f"FAIL: Component 2 — E2 is not a formula: {repr(e2_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check formula structure: {e}")

    # Component 3: Formula uses SUMPRODUCT aggregation and the
    # computed ROUND-then-SUM result equals 4,217,000 (0.4 points)
    # We verify this by:
    #   (a) Checking the formula contains SUMPRODUCT
    #   (b) Computing ROUND(salary, -3) for each row 2-51 and summing — must equal 4,217,000
    # We compute manually because openpyxl does not evaluate formulas.
    try:
        e2_value = ws['E2'].value
        has_sumproduct = (
            e2_value is not None
            and isinstance(e2_value, str)
            and 'SUMPRODUCT' in e2_value.upper().replace(' ', '')
        )

        # Compute the expected result manually from the salary data
        computed_total = 0
        salary_count = 0
        for row in range(2, 52):  # rows 2-51 inclusive (50 employees)
            salary = ws.cell(row=row, column=3).value  # column C
            if salary is not None:
                try:
                    rounded = round(float(salary) / 1000) * 1000
                    computed_total += rounded
                    salary_count += 1
                except (ValueError, TypeError):
                    pass

        expected_total = 4217000

        if has_sumproduct and salary_count == 50 and computed_total == expected_total:
            print(f"PASS: Component 3 — Formula uses SUMPRODUCT and computed total matches "
                  f"{expected_total} (salary_count={salary_count}, computed={computed_total}) (0.4 pts)")
            total_score += 0.4
        else:
            reasons = []
            if not has_sumproduct:
                reasons.append(f"formula does not use SUMPRODUCT (found: {repr(e2_value)})")
            if salary_count != 50:
                reasons.append(f"expected 50 salary rows, found {salary_count}")
            if computed_total != expected_total:
                reasons.append(f"computed ROUND-then-SUM = {computed_total}, expected {expected_total}")
            print(f"FAIL: Component 3 — {'; '.join(reasons)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
