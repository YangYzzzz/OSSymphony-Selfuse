"""
Reward Script: Invoice template with merged header, itemized line items,
subtotal/tax/total formulas, currency formatting, alternating row colors, and borders.
Task ID: calc_gsd_011
Domain: libreoffice_calc
Scoring:
  C1: Merged header A1:F1 with text/bold/size/center (0.20)
  C2: Row 4 column headers with bold (0.15)
  C3: Rows 5-14 contain line item data (0.15)
  C4: Rows 16-18 Subtotal/Tax/Total with formulas (0.20)
  C5: Currency format on monetary cells (0.10)
  C6: Alternating row colors rows 5-14 (0.10)
  C7: Thin borders on A4:F18 (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_011'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # ---------------------------------------------------------------
    # Component 1: Merged header A1:F1 with correct text, bold, 16pt,
    #              centered alignment (0.20 points)
    # ---------------------------------------------------------------
    try:
        # Check merge exists
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'F1' in r for r in merged_ranges)

        if not has_merge:
            print("FAIL: Component 1 -- A1:F1 not merged")
        else:
            c = ws['A1']
            val = c.value
            is_bold = c.font.bold is True
            font_size = c.font.size
            h_align = c.alignment.horizontal

            sub = 0.0
            if val and 'invoice' in str(val).lower():
                sub += 0.05
            else:
                print(f"FAIL: C1 sub -- header text missing or wrong: {val}")
            if is_bold:
                sub += 0.05
            else:
                print(f"FAIL: C1 sub -- header not bold")
            if font_size and float(font_size) >= 15:
                sub += 0.05
            else:
                print(f"FAIL: C1 sub -- font size {font_size}, expected ~16")
            if h_align == 'center':
                sub += 0.05
            else:
                print(f"FAIL: C1 sub -- alignment {h_align}, expected center")

            total_score += sub
            print(f"PASS: Component 1 -- merged header ({sub} pts)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # ---------------------------------------------------------------
    # Component 2: Row 4 column headers with bold formatting (0.15)
    # ---------------------------------------------------------------
    try:
        expected_headers = ['item', 'description', 'qty', 'unit price', 'discount', 'total']
        headers_found = 0
        bold_count = 0
        for col in range(1, 7):
            c = ws.cell(row=4, column=col)
            if c.value is not None:
                val_lower = str(c.value).strip().lower()
                # Check if it loosely matches any expected header
                for eh in expected_headers:
                    if eh in val_lower or val_lower in eh:
                        headers_found += 1
                        break
                if c.font.bold is True:
                    bold_count += 1

        sub = 0.0
        if headers_found >= 5:
            sub += 0.08
        else:
            print(f"FAIL: C2 sub -- only {headers_found}/6 headers found in row 4")
        if bold_count >= 5:
            sub += 0.07
        else:
            print(f"FAIL: C2 sub -- only {bold_count}/6 bold headers in row 4")

        total_score += sub
        print(f"PASS: Component 2 -- row 4 headers ({sub} pts)")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # ---------------------------------------------------------------
    # Component 3: Rows 5-14 contain line item data (0.15)
    # At least 8 out of 10 rows should have non-empty data in col A or B
    # ---------------------------------------------------------------
    try:
        rows_with_data = 0
        for row in range(5, 15):
            has_data = False
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                if c.value is not None and not isinstance(c, MergedCell):
                    has_data = True
                    break
            if has_data:
                rows_with_data += 1

        if rows_with_data >= 8:
            total_score += 0.15
            print(f"PASS: Component 3 -- {rows_with_data}/10 line item rows populated (0.15 pts)")
        elif rows_with_data >= 5:
            total_score += 0.08
            print(f"PARTIAL: Component 3 -- {rows_with_data}/10 rows (0.08 pts)")
        else:
            print(f"FAIL: Component 3 -- only {rows_with_data}/10 line item rows have data")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # ---------------------------------------------------------------
    # Component 4: Rows 16-18 Subtotal/Tax/Total with formulas (0.20)
    # ---------------------------------------------------------------
    try:
        sub = 0.0

        # Look for Subtotal, Tax, Total labels in rows 15-20
        # and corresponding formulas in column F
        subtotal_found = False
        tax_found = False
        total_found = False

        for row in range(15, 21):
            for col in range(1, 7):
                c = ws.cell(row=row, column=col)
                if c.value is not None:
                    val_str = str(c.value).strip().lower()
                    if 'subtotal' in val_str:
                        subtotal_found = True
                    elif 'tax' in val_str:
                        tax_found = True
                    elif val_str == 'total':
                        total_found = True

        # Check formulas in column F for those rows
        formula_count = 0
        for row in range(15, 21):
            c = ws.cell(row=row, column=6)
            if c.value is not None and isinstance(c.value, str) and c.value.startswith('='):
                formula_count += 1

        if subtotal_found:
            sub += 0.05
        else:
            print("FAIL: C4 sub -- 'Subtotal' label not found in rows 15-20")
        if tax_found:
            sub += 0.05
        else:
            print("FAIL: C4 sub -- 'Tax' label not found in rows 15-20")
        if total_found:
            sub += 0.05
        else:
            print("FAIL: C4 sub -- 'Total' label not found in rows 15-20")
        if formula_count >= 2:
            sub += 0.05
        else:
            print(f"FAIL: C4 sub -- only {formula_count} formulas in col F rows 15-20")

        total_score += sub
        print(f"PASS: Component 4 -- subtotal/tax/total section ({sub} pts)")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # ---------------------------------------------------------------
    # Component 5: Currency format on monetary cells (0.10)
    # Check D5, F5, F16, F17, F18 for dollar/currency format
    # ---------------------------------------------------------------
    try:
        currency_cells = []
        for coord in ['D5', 'F5', 'F16', 'F17', 'F18']:
            c = ws[coord]
            fmt = c.number_format if c.number_format else 'General'
            if '$' in fmt or 'USD' in fmt or '#,##0.00' in fmt:
                currency_cells.append(coord)

        if len(currency_cells) >= 4:
            total_score += 0.10
            print(f"PASS: Component 5 -- currency format on {len(currency_cells)}/5 cells (0.10 pts)")
        elif len(currency_cells) >= 2:
            total_score += 0.05
            print(f"PARTIAL: Component 5 -- currency format on {len(currency_cells)}/5 cells (0.05 pts)")
        else:
            print(f"FAIL: Component 5 -- currency format on only {len(currency_cells)}/5 cells")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # ---------------------------------------------------------------
    # Component 6: Alternating row colors on rows 5-14 (0.10)
    # Odd rows (5,7,9,11,13) should have light blue (#DDEEFF)
    # Even rows (6,8,10,12,14) should have white/no fill or different color
    # ---------------------------------------------------------------
    try:
        # First, count how many odd rows (5,7,9,11,13) have the blue fill.
        # If none have blue, the alternating pattern was never applied -> 0 pts.
        blue_rows = 0
        non_blue_rows = 0
        for row in range(5, 15):
            c = ws.cell(row=row, column=1)
            try:
                fill_type = c.fill.fill_type or c.fill.patternType
                rgb = c.fill.fgColor.rgb if fill_type else None
            except:
                rgb = None
                fill_type = None

            is_odd_row = (row % 2 == 1)  # rows 5,7,9,11,13
            has_blue = fill_type and rgb and 'DDEEFF' in str(rgb).upper()

            if is_odd_row and has_blue:
                blue_rows += 1
            elif not is_odd_row and not has_blue:
                non_blue_rows += 1

        # Gate: at least 3 odd rows must have blue fill to confirm the pattern exists
        if blue_rows < 3:
            print(f"FAIL: Component 6 -- only {blue_rows}/5 odd rows have blue fill; alternating pattern not applied")
        else:
            correct_alternating = blue_rows + non_blue_rows
            if correct_alternating >= 8:
                total_score += 0.10
                print(f"PASS: Component 6 -- {correct_alternating}/10 rows correctly colored (0.10 pts)")
            elif correct_alternating >= 5:
                total_score += 0.05
                print(f"PARTIAL: Component 6 -- {correct_alternating}/10 rows correct (0.05 pts)")
            else:
                print(f"FAIL: Component 6 -- only {correct_alternating}/10 rows have correct alternating colors")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # ---------------------------------------------------------------
    # Component 7: Thin borders on table A4:F18 (0.10)
    # Sample multiple cells to verify borders exist
    # ---------------------------------------------------------------
    try:
        border_cells_ok = 0
        check_coords = ['A4', 'F4', 'A10', 'F10', 'A18', 'F18', 'C8', 'D12']
        for coord in check_coords:
            c = ws[coord]
            b = c.border
            has_border = (
                b.left.style is not None or
                b.right.style is not None or
                b.top.style is not None or
                b.bottom.style is not None
            )
            if has_border:
                border_cells_ok += 1

        if border_cells_ok >= 6:
            total_score += 0.10
            print(f"PASS: Component 7 -- {border_cells_ok}/{len(check_coords)} cells have borders (0.10 pts)")
        elif border_cells_ok >= 3:
            total_score += 0.05
            print(f"PARTIAL: Component 7 -- {border_cells_ok}/{len(check_coords)} cells have borders (0.05 pts)")
        else:
            print(f"FAIL: Component 7 -- only {border_cells_ok}/{len(check_coords)} cells have borders")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
