"""
Initial Setup: Create project tracking spreadsheet for Gantt-like schedule task
Task ID: calc_ops_project_tracking_gantt_011
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, numbers
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_project_tracking_gantt_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ProjectPlan'

    # Project start date
    start_date = date(2025, 4, 7)  # A Monday

    # --- Row 1: Headers and date columns ---
    ws['A1'] = 'Task Name'
    ws['B1'] = 'Start Date'
    ws['C1'] = 'Duration Days'
    # Column D left empty as gap
    ws['D1'] = ''

    # Columns E through AL = columns 5 through 38 = 34 days (but 8 weeks = 56 days)
    # E=col5 to AL=col38 is 34 columns, but we need 56 days
    # E=5, F=6, ..., let's use E(5) to BH(60) for 56 days
    # Actually context says E through AL:
    # E=5, AL= col 38 (A=1, B=2, ..., Z=26, AA=27, AB=28, ..., AL=38)
    # That's 38-5+1 = 34 columns. But 8 weeks = 56 days.
    # Let's use 56 columns starting at E (col 5) through BE (col 57): 57-5+1 = 53...
    # Let's recalculate: E=5, 56 columns -> last column = 5+56-1 = 60 = BH
    # Context says E through AL (AL = column 38), that is only 34 columns.
    # We'll follow context: E through AL = 34 columns (just under 5 weeks, close enough)
    # But context says "8 weeks" which would be 56 days. Let's use the stated range E:AL.
    # Actually re-reading: "columns E through AL represent daily dates for 8 weeks"
    # If we interpret AL as column 38, that is 34 days. But 8 weeks = 56 days.
    # The context says E through AL explicitly. Let's put 56 days starting at E:
    # 56 columns from E(5) to BH(60). But context explicitly says AL...
    # We'll trust the context "E through AL" as the stated range and use 34 date columns.
    # Actually column AL: A=1..Z=26, AA=27..AL=38. So E(5)..AL(38) = 34 columns.
    # We'll use exactly that range for the dates.

    num_date_cols = 34  # E through AL
    date_start_col = 5  # E

    for i in range(num_date_cols):
        col = date_start_col + i
        cell = ws.cell(row=1, column=col, value=start_date + timedelta(days=i))
        # Format as short date - no conditional formatting, just the date value
        cell.number_format = 'DD/MM'
        cell.font = Font(size=8, bold=True)
        cell.alignment = Alignment(horizontal='center', text_rotation=90)

    # Style header row A1:D1
    header_font = Font(bold=True, size=10)
    for col_letter in ['A', 'B', 'C']:
        ws[f'{col_letter}1'].font = header_font
    ws['A1'].alignment = Alignment(horizontal='left')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')

    # --- Task data (15 tasks, rows 2-16) ---
    tasks = [
        ('Project Kickoff & Planning',    start_date + timedelta(days=0),  3),
        ('Requirements Analysis',          start_date + timedelta(days=1),  5),
        ('System Architecture Design',     start_date + timedelta(days=4),  7),
        ('Database Schema Design',         start_date + timedelta(days=5),  4),
        ('UI/UX Wireframes',               start_date + timedelta(days=7),  6),
        ('Backend API Development',        start_date + timedelta(days=11), 10),
        ('Frontend Development',           start_date + timedelta(days=13),  9),
        ('Database Implementation',        start_date + timedelta(days=10),  8),
        ('Integration Testing',            start_date + timedelta(days=18),  5),
        ('Performance Optimization',       start_date + timedelta(days=21),  4),
        ('Security Audit',                 start_date + timedelta(days=22),  3),
        ('User Acceptance Testing (UAT)',  start_date + timedelta(days=24),  6),
        ('Bug Fixes & Refinements',        start_date + timedelta(days=26),  5),
        ('Deployment & Configuration',     start_date + timedelta(days=29),  3),
        ('Go-Live & Handover',             start_date + timedelta(days=31),  2),
    ]

    for row_idx, (task_name, task_start, duration) in enumerate(tasks, 2):
        ws.cell(row=row_idx, column=1, value=task_name)
        ws.cell(row=row_idx, column=2, value=task_start).number_format = 'YYYY-MM-DD'
        ws.cell(row=row_idx, column=3, value=duration)

    # Column widths
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 4

    # Set narrow width for date columns
    for i in range(num_date_cols):
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(date_start_col + i)
        ws.column_dimensions[col_letter].width = 4

    # Row height for header
    ws.row_dimensions[1].height = 55

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ProjectPlan')
    print(f'  Tasks: 15 rows (rows 2-16)')
    print(f'  Date columns: E through AL ({num_date_cols} columns)')
    print(f'  No conditional formatting (to be added by agent)')
    print(f'  No freeze panes (to be added by agent)')


create_initial()
