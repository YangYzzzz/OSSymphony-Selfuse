"""
Initial Setup: Sales Commission Accelerator Spreadsheet
Task ID: calc_sales_commission_accelerator_007
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_commission_accelerator_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SalesComm'

    # Headers row 1
    headers = ['Rep', 'Quota', 'Total Sales', 'Attainment', 'Base Comm Rate',
               'Base Commission', 'Accelerator Bonus', 'Total Earnings']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 20 realistic sales reps with varied performance
    # Some above 110% quota, some between 100-110%, some below 100%
    reps_data = [
        ('Jordan Mitchell',    280000, 342000),   # 122.1% - above 110%
        ('Priya Sharma',       320000, 371200),   # 116.0% - above 110%
        ('Derek Wainwright',   250000, 287500),   # 115.0% - above 110%
        ('Aisha Okonkwo',      300000, 345000),   # 115.0% - above 110%
        ('Marcus Delgado',     270000, 308880),   # 114.4% - above 110%
        ('Natalie Bergström',  350000, 397250),   # 113.5% - above 110%
        ('Tyler Houghton',     240000, 270960),   # 112.9% - above 110%
        ('Lena Kovalenko',     290000, 325280),   # 112.2% - above 110%
        ('Samuel Osei',        260000, 289900),   # 111.5% - above 110%
        ('Fiona MacAllister',  310000, 344110),   # 111.0% - above 110%
        ('Carlos Venegas',     330000, 359700),   # 109.0% - below 110%
        ('Yuki Tanaka',        275000, 298375),   # 108.5% - below 110%
        ('Rachel Goldenberg',  295000, 318600),   # 108.0% - below 110%
        ('Ibrahim Al-Rashid',  260000, 278200),   # 107.0% - below 110%
        ('Sophie Laforest',    340000, 360400),   # 106.0% - below 110%
        ('Brandon Kessler',    285000, 296400),   # 104.0% - below 110%
        ('Miriam Oduya',       315000, 325575),   # 103.4% - below 110%
        ('Patrick Connolly',   270000, 275400),   # 102.0% - below 110%
        ('Anastasia Volkov',   290000, 287100),   # 98.9%  - below quota
        ('Wei Zhang',          325000, 314625),   # 96.8%  - below quota
    ]

    for r, (rep, quota, total_sales) in enumerate(reps_data, 2):
        base_rate = 0.08
        ws.cell(row=r, column=1, value=rep)
        ws.cell(row=r, column=2, value=quota).number_format = '#,##0'
        ws.cell(row=r, column=3, value=total_sales).number_format = '#,##0'
        # Attainment formula: =C/B
        ws.cell(row=r, column=4, value=f'=C{r}/B{r}').number_format = '0.0%'
        ws.cell(row=r, column=5, value=base_rate).number_format = '0%'
        # Base Commission formula: =C*E
        ws.cell(row=r, column=6, value=f'=C{r}*E{r}').number_format = '$#,##0.00'
        # Columns G (Accelerator Bonus) and H (Total Earnings) left EMPTY

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
