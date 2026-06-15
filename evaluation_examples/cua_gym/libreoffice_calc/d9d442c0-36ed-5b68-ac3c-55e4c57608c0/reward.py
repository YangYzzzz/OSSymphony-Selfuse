"""
Reward Script: Sales Commission Accelerator Model
Task ID: calc_sales_commission_accelerator_007
Domain: libreoffice_calc

Task: Add accelerator bonus column (G: IF attainment > 110% -> (Sales-Quota)*3%, else 0),
      total earnings column (H: base_comm + accelerator, currency format $#,##0.00),
      data sorted by Total Earnings descending, and a Rank column (I: RANK formula).

Scoring Rubric:
  Component 1: Accelerator Bonus formulas in G2:G21  — 0.30 pts
  Component 2: Total Earnings formulas in H2:H21     — 0.25 pts
  Component 3: Total Earnings currency format H2:H21 — 0.15 pts
  Component 4: Rank column I with RANK formula       — 0.20 pts
  Component 5: Data sorted by Total Earnings desc    — 0.10 pts
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_commission_accelerator_007'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: SalesComm sheet must exist
    if 'SalesComm' not in wb.sheetnames:
        print("CRITICAL: Sheet 'SalesComm' not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SalesComm']

    # Precondition: must have at least 21 rows (header + 20 reps)
    if ws.max_row < 21:
        print(f"CRITICAL: Expected at least 21 rows, found {ws.max_row}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Accelerator Bonus IF formula in G2:G21  (0.30 pts)
    # Expected formula pattern: =IF(Dx>1.10,(Cx-Bx)*0.03,0)
    # Must FAIL on initial (G is empty) and PASS on golden.
    # -----------------------------------------------------------------------
    try:
        g_formula_count = 0
        g_formula_correct = 0
        for row in range(2, 22):
            val = ws.cell(row=row, column=7).value
            if val is not None and isinstance(val, str):
                g_formula_count += 1
                # Check it's an IF formula with >1.10 threshold and *0.03 multiplier
                v_upper = val.upper().replace(' ', '')
                if 'IF(' in v_upper and '>1.10' in v_upper and '*0.03' in v_upper:
                    g_formula_correct += 1

        if g_formula_count == 20 and g_formula_correct == 20:
            print(f"PASS: Component 1 — All 20 accelerator bonus IF formulas present and correct in G2:G21 (0.30 pts)")
            total_score += 0.30
        elif g_formula_count == 20 and g_formula_correct > 0:
            # Partial: formulas present but not all with correct structure
            partial = round(0.15 * g_formula_correct / 20, 4)
            print(f"PARTIAL: Component 1 — {g_formula_correct}/20 formulas have correct IF(>1.10,*0.03) structure, {g_formula_count} formulas present. Partial: {partial} pts")
            total_score += partial
        elif g_formula_count > 0:
            print(f"PARTIAL: Component 1 — {g_formula_count}/20 cells in G have formulas, {g_formula_correct} with correct structure. Partial: 0.05 pts")
            total_score += 0.05
        else:
            print(f"FAIL: Component 1 — G2:G21 contains no formulas (all None or numeric). Expected IF accelerator formulas.")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Total Earnings formulas in H2:H21  (0.25 pts)
    # Expected: =Fx+Gx (combining Base Commission + Accelerator Bonus)
    # Must FAIL on initial (H is empty) and PASS on golden.
    # -----------------------------------------------------------------------
    try:
        h_formula_count = 0
        h_formula_correct = 0
        for row in range(2, 22):
            val = ws.cell(row=row, column=8).value
            if val is not None and isinstance(val, str):
                h_formula_count += 1
                # Check it's a formula combining F and G columns
                v_upper = val.upper().replace(' ', '')
                # Should reference both F and G column: =Fx+Gx or =Gx+Fx
                import re
                has_f = bool(re.search(r'F\d+', v_upper))
                has_g = bool(re.search(r'G\d+', v_upper))
                if has_f and has_g and '+' in v_upper:
                    h_formula_correct += 1

        if h_formula_count == 20 and h_formula_correct == 20:
            print(f"PASS: Component 2 — All 20 total earnings formulas (F+G) present in H2:H21 (0.25 pts)")
            total_score += 0.25
        elif h_formula_count == 20 and h_formula_correct > 0:
            partial = round(0.12 * h_formula_correct / 20, 4)
            print(f"PARTIAL: Component 2 — {h_formula_correct}/20 H formulas correctly reference F+G. Partial: {partial} pts")
            total_score += partial
        elif h_formula_count > 0:
            print(f"PARTIAL: Component 2 — {h_formula_count}/20 cells in H have formulas. Partial: 0.05 pts")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — H2:H21 contains no formulas. Expected total earnings =Fx+Gx formulas.")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Total Earnings currency format $#,##0.00 in H2:H21  (0.15 pts)
    # Must FAIL on initial (H empty, no format) and PASS on golden.
    # -----------------------------------------------------------------------
    try:
        h_currency_count = 0
        for row in range(2, 22):
            cell = ws.cell(row=row, column=8)
            fmt = cell.number_format
            # Accept $#,##0.00 or similar currency formats
            if fmt and ('$' in fmt or '0.00' in fmt) and '#' in fmt:
                h_currency_count += 1

        if h_currency_count == 20:
            print(f"PASS: Component 3 — All 20 H cells have currency format ($#,##0.00) (0.15 pts)")
            total_score += 0.15
        elif h_currency_count >= 15:
            partial = round(0.15 * h_currency_count / 20, 4)
            print(f"PARTIAL: Component 3 — {h_currency_count}/20 H cells have currency format. Partial: {partial} pts")
            total_score += partial
        elif h_currency_count > 0:
            print(f"PARTIAL: Component 3 — Only {h_currency_count}/20 H cells have currency format. Partial: 0.03 pts")
            total_score += 0.03
        else:
            # Check if any H cell has a format at all
            h_sample_fmt = ws.cell(row=2, column=8).number_format
            print(f"FAIL: Component 3 — H cells lack currency format. H2 format: {repr(h_sample_fmt)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Rank column I with RANK formula in I2:I21  (0.20 pts)
    # Header I1 = 'Rank', formulas reference H column with absolute range $H$2:$H$21
    # Must FAIL on initial (no column I) and PASS on golden.
    # -----------------------------------------------------------------------
    try:
        # Check column I header
        i_header = ws.cell(row=1, column=9).value
        i_rank_formulas = 0
        i_correct_rank = 0

        for row in range(2, 22):
            val = ws.cell(row=row, column=9).value
            if val is not None and isinstance(val, str):
                i_rank_formulas += 1
                v_upper = val.upper().replace(' ', '')
                # Should be a RANK formula referencing H column with absolute range
                if 'RANK(' in v_upper and 'H' in v_upper:
                    i_correct_rank += 1

        header_ok = i_header is not None and str(i_header).strip().lower() in ('rank', 'ranking')

        if header_ok and i_rank_formulas == 20 and i_correct_rank == 20:
            print(f"PASS: Component 4 — Rank header '{i_header}' present, all 20 RANK formulas in I2:I21 (0.20 pts)")
            total_score += 0.20
        elif i_rank_formulas == 20 and i_correct_rank == 20:
            # Missing header, but formulas correct
            print(f"PARTIAL: Component 4 — All 20 RANK formulas correct but header missing/wrong ('{i_header}'). Partial: 0.15 pts")
            total_score += 0.15
        elif i_rank_formulas > 0 and i_correct_rank > 0:
            partial = round(0.10 * i_correct_rank / 20, 4)
            print(f"PARTIAL: Component 4 — {i_correct_rank}/20 RANK formulas present. Partial: {partial} pts")
            total_score += partial
        elif i_rank_formulas > 0:
            print(f"PARTIAL: Component 4 — {i_rank_formulas} formulas in I but not RANK type. Partial: 0.03 pts")
            total_score += 0.03
        else:
            print(f"FAIL: Component 4 — Column I is empty or missing. Expected RANK formulas with header 'Rank'. I1={repr(i_header)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Data sorted by Total Earnings descending  (0.10 pts)
    # Since formulas are not evaluated by openpyxl, we verify that the
    # rows are in the correct descending order by computing Total Earnings manually
    # from B (Quota) and C (Total Sales) columns.
    # Total Earnings = F + G = C*0.08 + IF(C/B > 1.10, (C-B)*0.03, 0)
    # Must FAIL on initial (unsorted/no formula) and PASS on golden.
    # -----------------------------------------------------------------------
    try:
        earnings_list = []
        valid_rows = 0
        for row in range(2, 22):
            quota = ws.cell(row=row, column=2).value
            total_sales = ws.cell(row=row, column=3).value
            if quota is not None and total_sales is not None:
                try:
                    quota_f = float(quota)
                    sales_f = float(total_sales)
                    base_comm = sales_f * 0.08
                    attainment = sales_f / quota_f if quota_f > 0 else 0
                    accel_bonus = (sales_f - quota_f) * 0.03 if attainment > 1.10 else 0
                    total_earnings = base_comm + accel_bonus
                    earnings_list.append(total_earnings)
                    valid_rows += 1
                except (TypeError, ValueError):
                    earnings_list.append(None)
            else:
                earnings_list.append(None)

        if valid_rows >= 20:
            # Check descending order
            is_sorted = True
            for idx in range(len(earnings_list) - 1):
                if earnings_list[idx] is not None and earnings_list[idx + 1] is not None:
                    if earnings_list[idx] < earnings_list[idx + 1]:
                        is_sorted = False
                        break
            if is_sorted:
                print(f"PASS: Component 5 — Data correctly sorted by Total Earnings descending (0.10 pts)")
                total_score += 0.10
            else:
                # Check if at least top 5 are sorted correctly
                top_sorted = all(
                    earnings_list[i] >= earnings_list[i + 1]
                    for i in range(min(4, len(earnings_list) - 1))
                    if earnings_list[i] is not None and earnings_list[i + 1] is not None
                )
                if top_sorted:
                    print(f"PARTIAL: Component 5 — Top rows sorted correctly but overall order fails. Partial: 0.05 pts")
                    total_score += 0.05
                else:
                    first_three = [round(e, 2) if e else None for e in earnings_list[:3]]
                    print(f"FAIL: Component 5 — Data NOT sorted by Total Earnings descending. Top 3 earnings: {first_three}")
        else:
            print(f"FAIL: Component 5 — Could not read enough valid rows ({valid_rows}/20) to verify sorting.")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
