"""
Initial Setup: Sort the sales pipeline by deal value from highest to lowest.
Task ID: calc_sales_001
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Pipeline'

    # Headers
    headers = ['Deal Name', 'Account', 'Stage', 'Deal Value']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Data rows - UNSORTED original order
    data = [
        ['Acme Expansion', 'Acme Corp', 'Negotiation', 45000],
        ['Beta Renewal', 'Beta Inc', 'Closed Won', 120000],
        ['Gamma Upsell', 'Gamma LLC', 'Proposal', 78000],
        ['Delta New', 'Delta Co', 'Discovery', 32000],
        ['Epsilon Deal', 'Epsilon Ltd', 'Negotiation', 95000],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
