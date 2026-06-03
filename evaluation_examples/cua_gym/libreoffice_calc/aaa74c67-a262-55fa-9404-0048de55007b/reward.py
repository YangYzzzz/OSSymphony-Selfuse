"""
Reward Script: Sort the sales pipeline by deal value from highest to lowest.
Task ID: calc_sales_001
Domain: libreoffice_calc
Scoring: 3 components checking sort order, top row, and bottom row
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_001'

# Ground truth: rows sorted descending by Deal Value (column D)
EXPECTED_ORDER = [
    ("Beta Renewal", 120000),
    ("Epsilon Deal", 95000),
    ("Gamma Upsell", 78000),
    ("Acme Expansion", 45000),
    ("Delta New", 32000),
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Pipeline sheet exists
    if 'Pipeline' not in wb.sheetnames:
        print("CRITICAL: 'Pipeline' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pipeline']

    # Read actual data rows (2-6)
    actual_rows = []
    for r in range(2, 7):
        name = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=4).value
        actual_rows.append((str(name).strip() if name else None, value))

    # Component 1: Data is sorted in descending order by Deal Value (0.4 points)
    # Check that column D values are strictly descending from row 2 to row 6
    try:
        values = []
        for name, val in actual_rows:
            if val is not None:
                values.append(float(val))
            else:
                values.append(None)

        is_descending = all(
            values[i] is not None and values[i+1] is not None and values[i] > values[i+1]
            for i in range(len(values) - 1)
        )

        if is_descending:
            print(f"PASS: Component 1 -- Deal Values are in strict descending order: {values} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 -- Deal Values are NOT in descending order: {values}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Top two rows are correctly positioned (highest values) (0.3 points)
    # Row 2 must be Beta Renewal (120000) AND Row 3 must be Epsilon Deal (95000)
    # These rows are NOT in these positions in the initial file, so this truly measures sorting.
    try:
        row2_name, row2_val = actual_rows[0]
        row3_name, row3_val = actual_rows[1]

        top_two_correct = (
            row2_name == "Beta Renewal" and abs(float(row2_val) - 120000) < 0.01 and
            row3_name == "Epsilon Deal" and abs(float(row3_val) - 95000) < 0.01
        )

        if top_two_correct:
            print(f"PASS: Component 2 -- Top two rows correct: '{row2_name}' ({row2_val}), '{row3_name}' ({row3_val}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 -- Top two rows incorrect: '{row2_name}' ({row2_val}), '{row3_name}' ({row3_val})")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Bottom two rows are correctly positioned (lowest values) (0.3 points)
    # Row 5 must be Acme Expansion (45000) AND Row 6 must be Delta New (32000)
    # In initial file, row 5 has Delta New and row 6 has Epsilon Deal, so neither match.
    try:
        row5_name, row5_val = actual_rows[3]
        row6_name, row6_val = actual_rows[4]

        bottom_two_correct = (
            row5_name == "Acme Expansion" and abs(float(row5_val) - 45000) < 0.01 and
            row6_name == "Delta New" and abs(float(row6_val) - 32000) < 0.01
        )

        if bottom_two_correct:
            print(f"PASS: Component 3 -- Bottom two rows correct: '{row5_name}' ({row5_val}), '{row6_name}' ({row6_val}) (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 -- Bottom two rows incorrect: '{row5_name}' ({row5_val}), '{row6_name}' ({row6_val})")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
