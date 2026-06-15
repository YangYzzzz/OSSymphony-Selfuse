"""
Initial Setup: Set up the 'Inventory' sheet for printing
Task ID: calc_mcp_074
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # --- Headers in row 1 ---
    headers = [
        'Item Code', 'Product Name', 'Category', 'Warehouse',
        'Qty On Hand', 'Reorder Level', 'Unit Cost', 'Total Value'
    ]
    header_font = Font(name='Calibri', size=11, bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # --- Data generation for rows 2-100 (99 data rows) ---
    categories = ['Electronics', 'Furniture', 'Office Supplies', 'Clothing',
                  'Food & Beverage', 'Hardware', 'Cleaning', 'Safety Equipment']
    warehouses = ['WH-North', 'WH-South', 'WH-East', 'WH-West', 'WH-Central']

    products = [
        ('EL-1001', 'Wireless Mouse', 'Electronics'),
        ('EL-1002', 'USB-C Hub', 'Electronics'),
        ('EL-1003', 'Monitor Stand', 'Electronics'),
        ('EL-1004', '27" LED Monitor', 'Electronics'),
        ('EL-1005', 'Mechanical Keyboard', 'Electronics'),
        ('EL-1006', 'Webcam HD Pro', 'Electronics'),
        ('EL-1007', 'Noise-Cancel Headset', 'Electronics'),
        ('EL-1008', 'Portable SSD 1TB', 'Electronics'),
        ('FU-2001', 'Adjustable Desk', 'Furniture'),
        ('FU-2002', 'Ergonomic Chair', 'Furniture'),
        ('FU-2003', 'File Cabinet 4-Drawer', 'Furniture'),
        ('FU-2004', 'Bookshelf Tall', 'Furniture'),
        ('FU-2005', 'Meeting Table 8-Seat', 'Furniture'),
        ('FU-2006', 'Storage Locker', 'Furniture'),
        ('OS-3001', 'Copy Paper A4 500ct', 'Office Supplies'),
        ('OS-3002', 'Ballpoint Pen Box 50', 'Office Supplies'),
        ('OS-3003', 'Sticky Notes Bulk', 'Office Supplies'),
        ('OS-3004', 'Binder Clips Assorted', 'Office Supplies'),
        ('OS-3005', 'Whiteboard Markers 12pk', 'Office Supplies'),
        ('OS-3006', 'Desk Organizer', 'Office Supplies'),
        ('OS-3007', 'Laminating Pouches 100pk', 'Office Supplies'),
        ('OS-3008', 'Stapler Heavy Duty', 'Office Supplies'),
        ('CL-4001', 'Safety Vest Hi-Vis', 'Clothing'),
        ('CL-4002', 'Work Gloves Leather', 'Clothing'),
        ('CL-4003', 'Steel Toe Boots', 'Clothing'),
        ('CL-4004', 'Lab Coat White', 'Clothing'),
        ('CL-4005', 'Rain Jacket', 'Clothing'),
        ('FB-5001', 'Coffee Beans 5lb', 'Food & Beverage'),
        ('FB-5002', 'Bottled Water 24pk', 'Food & Beverage'),
        ('FB-5003', 'Tea Assortment Box', 'Food & Beverage'),
        ('FB-5004', 'Sugar Packets 1000ct', 'Food & Beverage'),
        ('FB-5005', 'Creamer Pods 200ct', 'Food & Beverage'),
        ('HW-6001', 'Drill Bit Set', 'Hardware'),
        ('HW-6002', 'Cable Ties 500pk', 'Hardware'),
        ('HW-6003', 'Screwdriver Multi-Set', 'Hardware'),
        ('HW-6004', 'Wire Stripper', 'Hardware'),
        ('HW-6005', 'Electrical Tape 10pk', 'Hardware'),
        ('HW-6006', 'Socket Wrench Set', 'Hardware'),
        ('CG-7001', 'All-Purpose Cleaner 1gal', 'Cleaning'),
        ('CG-7002', 'Microfiber Cloths 50pk', 'Cleaning'),
        ('CG-7003', 'Floor Polish', 'Cleaning'),
        ('CG-7004', 'Hand Sanitizer 1L', 'Cleaning'),
        ('CG-7005', 'Trash Bags 100ct', 'Cleaning'),
        ('SE-8001', 'Fire Extinguisher', 'Safety Equipment'),
        ('SE-8002', 'First Aid Kit', 'Safety Equipment'),
        ('SE-8003', 'Safety Goggles', 'Safety Equipment'),
        ('SE-8004', 'Ear Plugs 200pk', 'Safety Equipment'),
        ('SE-8005', 'Hard Hat', 'Safety Equipment'),
        ('SE-8006', 'Spill Kit', 'Safety Equipment'),
    ]

    random.seed(42)

    # Fill 99 data rows (rows 2..100)
    for r in range(2, 101):
        idx = (r - 2) % len(products)
        item_code, product_name, category = products[idx]
        # Add a suffix for repeating products to keep them unique
        cycle = (r - 2) // len(products)
        if cycle > 0:
            item_code = f"{item_code}-{cycle}"

        warehouse = warehouses[(r - 2) % len(warehouses)]
        qty = random.randint(5, 500)
        reorder = random.randint(10, 100)
        unit_cost = round(random.uniform(2.50, 450.00), 2)
        total_value = round(qty * unit_cost, 2)

        ws.cell(row=r, column=1, value=item_code)
        ws.cell(row=r, column=2, value=product_name)
        ws.cell(row=r, column=3, value=category)
        ws.cell(row=r, column=4, value=warehouse)
        ws.cell(row=r, column=5, value=qty)
        ws.cell(row=r, column=6, value=reorder)
        ws.cell(row=r, column=7, value=unit_cost)
        ws.cell(row=r, column=8, value=total_value)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 26
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14

    # NO print area, NO repeat columns, NO landscape, NO scale — these are the task
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
