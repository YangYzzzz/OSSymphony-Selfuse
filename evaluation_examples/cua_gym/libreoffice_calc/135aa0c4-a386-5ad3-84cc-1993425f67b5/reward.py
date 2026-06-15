"""
Reward Script: Set up Inventory sheet for printing
Task ID: calc_mcp_074
Domain: libreoffice_calc
Scoring:
  Component 1: Print area = A1:H100 (0.3 pts)
  Component 2: Repeat columns A:B on every page (0.3 pts)
  Component 3: Landscape orientation (0.2 pts)
  Component 4: Scale = 90% (0.2 pts)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_074'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice session."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify print setup on the Inventory sheet with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Inventory sheet must exist
    if 'Inventory' not in wb.sheetnames:
        print(f"FAIL: Sheet 'Inventory' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']
    ps = ws.page_setup

    # Component 1: Print area = A1:H100 (0.3 points)
    try:
        print_area = ws.print_area
        # print_area can be a string like "'Inventory'!$A$1:$H$100" or "$A$1:$H$100"
        # Normalize: check that it references A1:H100
        if print_area:
            pa_str = str(print_area).upper()
            # Strip sheet name prefix if present
            if '!' in pa_str:
                pa_str = pa_str.split('!')[-1]
            # Remove $ signs for easier comparison
            pa_normalized = pa_str.replace('$', '').strip()
            if pa_normalized == 'A1:H100':
                print(f"PASS: Component 1 - Print area is A1:H100 (raw: {print_area}) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 1 - Print area normalized to '{pa_normalized}', expected 'A1:H100' (raw: {print_area})")
        else:
            print(f"FAIL: Component 1 - No print area set (value: {print_area})")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Repeat columns A:B on every page (0.3 points)
    try:
        title_cols = ws.print_title_cols
        if title_cols:
            tc_str = str(title_cols).upper().replace('$', '').strip()
            if tc_str == 'A:B':
                print(f"PASS: Component 2 - Repeat columns set to A:B (raw: {title_cols}) (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 - Repeat columns normalized to '{tc_str}', expected 'A:B' (raw: {title_cols})")
        else:
            print(f"FAIL: Component 2 - No repeat columns set (value: {title_cols})")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Landscape orientation (0.2 points)
    try:
        orientation = ps.orientation
        if orientation and str(orientation).lower() == 'landscape':
            print(f"PASS: Component 3 - Orientation is landscape (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 - Orientation is '{orientation}', expected 'landscape'")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Scale = 90% (0.2 points)
    try:
        scale = ps.scale
        if scale is not None:
            scale_val = int(scale)
            if scale_val == 90:
                print(f"PASS: Component 4 - Scale is 90% (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 4 - Scale is {scale_val}%, expected 90%")
        else:
            print(f"FAIL: Component 4 - Scale not set (value: {scale})")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
