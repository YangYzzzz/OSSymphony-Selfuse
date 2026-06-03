"""
Reward Script: Build a dynamic org chart data model with INDIRECT formulas
Task ID: calc_hr_051
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): C2 has COUNTA+INDIRECT formula referencing Eng sheet
  Component 2 (0.3): C3 has COUNTA+INDIRECT formula referencing Sales sheet
  Component 3 (0.3): C4 has COUNTA+INDIRECT formula referencing HR sheet
  Component 4 (0.1): All formulas dynamically reference column B for sheet name
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_051'


def normalize_formula(f):
    """Normalize a formula string for comparison: uppercase, strip spaces."""
    if not isinstance(f, str):
        return ""
    return f.upper().replace(" ", "")


def is_indirect_counta_formula(formula_str, expected_row):
    """
    Check if formula is a COUNTA(INDIRECT(...)) formula that dynamically
    references a sheet using the cell in column B of the expected row.

    Accepts variations like:
      =COUNTA(INDIRECT(B2&".A2:A100"))
      =COUNTA(INDIRECT(B2&"!A2:A100"))
      =COUNTA(INDIRECT(B2&".A:A"))
      etc.
    """
    norm = normalize_formula(formula_str)
    if not norm.startswith("="):
        return False

    # Must contain both COUNTA and INDIRECT
    if "COUNTA" not in norm or "INDIRECT" not in norm:
        return False

    # Must reference B{expected_row} dynamically (the sheet name cell)
    b_ref = f"B{expected_row}"
    if b_ref not in norm:
        return False

    return True


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Consolidated sheet must exist
    if 'Consolidated' not in wb.sheetnames:
        print("FAIL: 'Consolidated' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Consolidated']

    # Component 1: C2 has COUNTA+INDIRECT formula referencing Eng sheet (0.3 points)
    try:
        c2_val = ws['C2'].value
        if c2_val and is_indirect_counta_formula(str(c2_val), 2):
            print(f"PASS: Component 1 — C2 has valid INDIRECT formula: {c2_val} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — C2 expected COUNTA(INDIRECT(B2&...)) formula, found: {repr(c2_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C3 has COUNTA+INDIRECT formula referencing Sales sheet (0.3 points)
    try:
        c3_val = ws['C3'].value
        if c3_val and is_indirect_counta_formula(str(c3_val), 3):
            print(f"PASS: Component 2 — C3 has valid INDIRECT formula: {c3_val} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — C3 expected COUNTA(INDIRECT(B3&...)) formula, found: {repr(c3_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: C4 has COUNTA+INDIRECT formula referencing HR sheet (0.3 points)
    try:
        c4_val = ws['C4'].value
        if c4_val and is_indirect_counta_formula(str(c4_val), 4):
            print(f"PASS: Component 3 — C4 has valid INDIRECT formula: {c4_val} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — C4 expected COUNTA(INDIRECT(B4&...)) formula, found: {repr(c4_val)}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: All three formulas use dynamic sheet references via column B (0.1 points)
    # This checks that the formulas form a consistent pattern — all three use INDIRECT
    # with their respective B-column cell, ensuring the model is truly dynamic.
    try:
        dynamic_count = sum(
            1 for row_num in [2, 3, 4]
            if ws.cell(row=row_num, column=3).value
            and is_indirect_counta_formula(str(ws.cell(row=row_num, column=3).value), row_num)
        )

        if dynamic_count == 3:
            # Additionally verify B column has sheet names that match actual sheets
            b2 = ws['B2'].value
            b3 = ws['B3'].value
            b4 = ws['B4'].value
            sheets_match = (
                b2 in wb.sheetnames and
                b3 in wb.sheetnames and
                b4 in wb.sheetnames
            )
            if sheets_match:
                print(f"PASS: Component 4 — All formulas dynamically reference column B sheet names (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 4 — B column values {[b2, b3, b4]} don't all match sheet names {wb.sheetnames}")
        else:
            print(f"FAIL: Component 4 — Only {dynamic_count}/3 C2:C4 cells have dynamic INDIRECT formulas")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
