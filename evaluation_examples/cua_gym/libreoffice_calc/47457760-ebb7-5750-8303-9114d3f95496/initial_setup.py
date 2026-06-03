"""
Initial Setup: Build org chart data model with department sheets and consolidated report
Task ID: calc_hr_051
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Eng (Engineering) ---
    ws_eng = wb.active
    ws_eng.title = 'Eng'
    ws_eng['A1'] = 'Name'
    ws_eng['B1'] = 'Title'
    eng_data = [
        ['Sarah Chen', 'Senior Software Engineer'],
        ['Marcus Johnson', 'Backend Developer'],
        ['Priya Patel', 'DevOps Engineer'],
        ['James O\'Brien', 'Frontend Developer'],
        ['Lin Wei', 'QA Lead'],
    ]
    for r, row in enumerate(eng_data, 2):
        ws_eng.cell(row=r, column=1, value=row[0])
        ws_eng.cell(row=r, column=2, value=row[1])

    # --- Sheet 2: Sales ---
    ws_sales = wb.create_sheet('Sales')
    ws_sales['A1'] = 'Name'
    ws_sales['B1'] = 'Title'
    sales_data = [
        ['Diana Rodriguez', 'Regional Sales Manager'],
        ['Thomas Kim', 'Account Executive'],
        ['Rachel Adams', 'Sales Development Rep'],
        ['Omar Hassan', 'Enterprise Account Manager'],
    ]
    for r, row in enumerate(sales_data, 2):
        ws_sales.cell(row=r, column=1, value=row[0])
        ws_sales.cell(row=r, column=2, value=row[1])

    # --- Sheet 3: HR ---
    ws_hr = wb.create_sheet('HR')
    ws_hr['A1'] = 'Name'
    ws_hr['B1'] = 'Title'
    hr_data = [
        ['Emily Foster', 'HR Director'],
        ['Carlos Mendez', 'Recruiter'],
        ['Aisha Williams', 'Benefits Coordinator'],
    ]
    for r, row in enumerate(hr_data, 2):
        ws_hr.cell(row=r, column=1, value=row[0])
        ws_hr.cell(row=r, column=2, value=row[1])

    # --- Sheet 4: Consolidated ---
    ws_con = wb.create_sheet('Consolidated')
    ws_con['A1'] = 'Department'
    ws_con['B1'] = 'Sheet Name'
    ws_con['C1'] = 'Headcount'

    ws_con['A2'] = 'Engineering'
    ws_con['B2'] = 'Eng'
    # C2 left empty - agent must enter INDIRECT formula

    ws_con['A3'] = 'Sales'
    ws_con['B3'] = 'Sales'
    # C3 left empty

    ws_con['A4'] = 'HR'
    ws_con['B4'] = 'HR'
    # C4 left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
