"""
Reward Script: Fleet Fuel Efficiency Tracker
Task ID: calc_wf_091
Domain: libreoffice_calc
Scoring:
  1. MPG calculation formulas in Fuel Log col G (0.20)
  2. Rolling average MPG formulas in Fuel Log col H (0.15)
  3. Summary sheet with per-vehicle stats (0.20)
  4. Fleet avg MPG row and Flag column with CHECK logic (0.15)
  5. Conditional formatting (red for flagged, orange for outliers) (0.10)
  6. Charts exist (line + bar) (0.10)
  7. Chart data correctness (0.10)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_091'


def verify_task(file_path):
    """Verify task completion with progressive scoring. Returns float 0.0-1.0."""
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # --- Precondition: Fuel Log sheet must exist ---
    if 'Fuel Log' not in wb.sheetnames:
        print("CRITICAL: 'Fuel Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_fuel = wb['Fuel Log']

    # Determine vehicle boundaries (each vehicle has 6 rows of data)
    # Vehicles start at rows 2,8,14,20,26,32,38,44,50,56
    vehicle_starts = []
    prev_vid = None
    for r in range(2, ws_fuel.max_row + 1):
        vid = ws_fuel.cell(r, 2).value
        if vid is not None and vid != prev_vid:
            vehicle_starts.append(r)
            prev_vid = vid

    # ---------------------------------------------------------------
    # Component 1: MPG calculation formulas in col G (0.20 points)
    # ---------------------------------------------------------------
    try:
        mpg_formula_count = 0
        mpg_expected = 0
        for idx, start_row in enumerate(vehicle_starts):
            # Each vehicle has 6 entries; first entry has no MPG, entries 2-6 have MPG
            end_row = vehicle_starts[idx + 1] - 1 if idx + 1 < len(vehicle_starts) else ws_fuel.max_row
            for r in range(start_row + 1, end_row + 1):
                mpg_expected += 1
                val = ws_fuel.cell(r, 7).value  # Column G
                if isinstance(val, str) and val.startswith('='):
                    # Check it references odometer diff / gallons
                    upper = val.upper().replace(' ', '')
                    if 'C' in upper and 'D' in upper and '/' in upper:
                        mpg_formula_count += 1

        if mpg_expected > 0 and mpg_formula_count >= mpg_expected * 0.8:
            pts = 0.20
            print(f"PASS: Component 1 - MPG formulas found ({mpg_formula_count}/{mpg_expected}) (0.20 pts)")
            total_score += pts
        elif mpg_formula_count > 0:
            pts = 0.10
            print(f"PARTIAL: Component 1 - Some MPG formulas ({mpg_formula_count}/{mpg_expected}) (0.10 pts)")
            total_score += pts
        else:
            print(f"FAIL: Component 1 - No MPG formulas in col G")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # ---------------------------------------------------------------
    # Component 2: Rolling average MPG formulas in col H (0.15 points)
    # ---------------------------------------------------------------
    try:
        rolling_count = 0
        rolling_expected = 0
        for idx, start_row in enumerate(vehicle_starts):
            end_row = vehicle_starts[idx + 1] - 1 if idx + 1 < len(vehicle_starts) else ws_fuel.max_row
            for r in range(start_row + 1, end_row + 1):
                rolling_expected += 1
                val = ws_fuel.cell(r, 8).value  # Column H
                if isinstance(val, str) and val.startswith('='):
                    upper = val.upper().replace(' ', '')
                    if 'AVERAGE' in upper and 'G' in upper:
                        rolling_count += 1

        if rolling_expected > 0 and rolling_count >= rolling_expected * 0.8:
            print(f"PASS: Component 2 - Rolling avg MPG formulas ({rolling_count}/{rolling_expected}) (0.15 pts)")
            total_score += 0.15
        elif rolling_count > 0:
            print(f"PARTIAL: Component 2 - Some rolling avg formulas ({rolling_count}/{rolling_expected}) (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2 - No rolling avg MPG formulas in col H")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # ---------------------------------------------------------------
    # Component 3: Summary sheet with per-vehicle stats (0.20 points)
    # ---------------------------------------------------------------
    try:
        if 'Summary' in wb.sheetnames:
            ws_sum = wb['Summary']
            # Check headers
            expected_headers = {'vehicle', 'total miles', 'total gallons', 'avg mpg',
                                'total cost', 'cost per mile', 'flag'}
            actual_headers = set()
            for c in range(1, ws_sum.max_column + 1):
                h = ws_sum.cell(1, c).value
                if h:
                    actual_headers.add(str(h).lower().strip())

            header_match = len(expected_headers.intersection(actual_headers)) >= 5

            # Check that at least 8 vehicles have formula-based stats
            formula_rows = 0
            for r in range(2, min(ws_sum.max_row + 1, 15)):
                vid = ws_sum.cell(r, 1).value
                total_miles = ws_sum.cell(r, 2).value
                if vid and isinstance(total_miles, str) and total_miles.startswith('='):
                    formula_rows += 1

            if header_match and formula_rows >= 8:
                print(f"PASS: Component 3 - Summary sheet with {formula_rows} vehicle rows (0.20 pts)")
                total_score += 0.20
            elif formula_rows >= 4:
                print(f"PARTIAL: Component 3 - Summary has {formula_rows} vehicles (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 3 - Summary sheet missing or incomplete (headers={header_match}, rows={formula_rows})")
        else:
            print("FAIL: Component 3 - 'Summary' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # ---------------------------------------------------------------
    # Component 4: Fleet avg MPG and Flag column (0.15 points)
    # ---------------------------------------------------------------
    try:
        if 'Summary' in wb.sheetnames:
            ws_sum = wb['Summary']
            fleet_avg_count = 0
            flag_formula_count = 0

            # Look for fleet average row (typically last data row)
            for r in range(2, ws_sum.max_row + 1):
                cell_a = ws_sum.cell(r, 1).value
                if cell_a and 'fleet' in str(cell_a).lower():
                    # Check it has a formula for avg MPG
                    avg_mpg_val = ws_sum.cell(r, 4).value
                    if isinstance(avg_mpg_val, str) and avg_mpg_val.startswith('='):
                        fleet_avg_count += 1

            # Check Flag column formulas (col G for vehicle rows)
            for r in range(2, ws_sum.max_row + 1):
                vid = ws_sum.cell(r, 1).value
                if vid and 'fleet' not in str(vid).lower():
                    flag_val = ws_sum.cell(r, 7).value
                    if isinstance(flag_val, str) and flag_val.startswith('='):
                        upper = flag_val.upper().replace(' ', '')
                        if 'IF' in upper and 'CHECK' in upper:
                            flag_formula_count += 1

            if fleet_avg_count >= 1 and flag_formula_count >= 8:
                print(f"PASS: Component 4 - Fleet avg found, {flag_formula_count} flag formulas (0.15 pts)")
                total_score += 0.15
            elif fleet_avg_count >= 1 or flag_formula_count >= 4:
                pts = 0.05 + (0.05 if fleet_avg_count >= 1 else 0) + (0.05 if flag_formula_count >= 4 else 0)
                pts = min(pts, 0.10)
                print(f"PARTIAL: Component 4 - fleet_avg={fleet_avg_count}, flags={flag_formula_count} ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 4 - No fleet avg or flag formulas")
        else:
            print("FAIL: Component 4 - 'Summary' sheet does not exist")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # ---------------------------------------------------------------
    # Component 5: Conditional formatting (0.10 points)
    # ---------------------------------------------------------------
    try:
        cf_outlier_count = 0
        cf_flagged_count = 0

        # Check Fuel Log for orange conditional formatting on gallons
        cf_fuel = ws_fuel.conditional_formatting
        for cf_range in cf_fuel:
            for rule in cf_range.rules:
                if rule.formula:
                    formula_str = str(rule.formula).upper()
                    if 'STDEV' in formula_str or 'STD' in formula_str:
                        cf_outlier_count += 1

        # Check Summary sheet for red conditional formatting on flagged
        if 'Summary' in wb.sheetnames:
            ws_sum = wb['Summary']
            cf_sum = ws_sum.conditional_formatting
            for cf_range in cf_sum:
                for rule in cf_range.rules:
                    if rule.formula:
                        formula_str = str(rule.formula).upper()
                        if 'CHECK' in formula_str:
                            cf_flagged_count += 1

        if cf_outlier_count >= 1 and cf_flagged_count >= 1:
            print(f"PASS: Component 5 - Both conditional formats found (0.10 pts)")
            total_score += 0.10
        elif cf_outlier_count >= 1 or cf_flagged_count >= 1:
            print(f"PARTIAL: Component 5 - outlier_rules={cf_outlier_count}, flagged_rules={cf_flagged_count} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 - No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # ---------------------------------------------------------------
    # Component 6: Charts exist - line chart + bar chart (0.10 points)
    # ---------------------------------------------------------------
    try:
        all_charts = []
        for sname in wb.sheetnames:
            ws_temp = wb[sname]
            if hasattr(ws_temp, '_charts'):
                all_charts.extend(ws_temp._charts)

        line_count = sum(1 for ch in all_charts if 'line' in getattr(ch, 'tagname', '').lower())
        bar_count = sum(1 for ch in all_charts if 'bar' in getattr(ch, 'tagname', '').lower())

        if line_count >= 1 and bar_count >= 1:
            print(f"PASS: Component 6 - Line chart and bar chart found (0.10 pts)")
            total_score += 0.10
        elif line_count >= 1 or bar_count >= 1:
            print(f"PARTIAL: Component 6 - line={line_count}, bar={bar_count} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 - No charts found (total charts: {len(all_charts)})")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # ---------------------------------------------------------------
    # Component 7: Chart data correctness (0.10 points)
    # ---------------------------------------------------------------
    try:
        line_series_count = 0
        bar_series_count = 0

        for sname in wb.sheetnames:
            ws_temp = wb[sname]
            if not hasattr(ws_temp, '_charts'):
                continue
            for ch in ws_temp._charts:
                tagname = getattr(ch, 'tagname', '')
                if 'line' in tagname.lower():
                    # Line chart should have multiple series (one per vehicle)
                    line_series_count = len(ch.series)
                if 'bar' in tagname.lower():
                    bar_series_count = len(ch.series)

        line_ok = line_series_count >= 5
        bar_ok = bar_series_count >= 1

        if line_ok and bar_ok:
            print(f"PASS: Component 7 - Line has {line_series_count} series, bar has {bar_series_count} series (0.10 pts)")
            total_score += 0.10
        elif line_ok or bar_ok:
            print(f"PARTIAL: Component 7 - line_series={line_series_count}, bar_series={bar_series_count} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 - Chart data references incorrect (line={line_series_count}, bar={bar_series_count})")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
