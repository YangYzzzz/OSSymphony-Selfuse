#!/usr/bin/env python3
"""initial_setup.py - Create the initial fuel log workbook with raw data only."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, subprocess, shlex, time

OUTPUT_PATH = "/home/user/calc_wf_091.xlsx"

# ── Shared raw data (10 vehicles × 6 fill-ups = 60 rows) ──────────────────
# Each tuple: (Date, Vehicle ID, Odometer, Gallons, Price/Gallon, Total Cost)
# Sorted by Vehicle ID then Date.  Odometer increases per vehicle.
RAW_DATA = [
    # V001 – sedan, decent mileage
    ("2025-01-05", "V001", 12000, 12.5, 3.45, 43.13),
    ("2025-01-19", "V001", 12380, 11.8, 3.52, 41.54),
    ("2025-02-02", "V001", 12740, 12.1, 3.49, 42.23),
    ("2025-02-16", "V001", 13110, 11.5, 3.55, 40.83),
    ("2025-03-02", "V001", 13500, 12.3, 3.60, 44.28),
    ("2025-03-16", "V001", 13870, 11.9, 3.58, 42.60),
    # V002 – truck, lower mileage
    ("2025-01-03", "V002", 45000, 16.2, 3.89, 63.02),
    ("2025-01-17", "V002", 45260, 15.8, 3.92, 61.94),
    ("2025-02-01", "V002", 45510, 16.5, 3.85, 63.53),
    ("2025-02-15", "V002", 45780, 16.0, 3.90, 62.40),
    ("2025-03-01", "V002", 46030, 15.5, 3.95, 61.23),
    ("2025-03-15", "V002", 46300, 16.3, 3.88, 63.24),
    # V003 – hybrid, high mileage
    ("2025-01-06", "V003", 8000, 8.5, 3.42, 29.07),
    ("2025-01-20", "V003", 8420, 8.2, 3.48, 28.54),
    ("2025-02-03", "V003", 8830, 8.8, 3.45, 30.36),
    ("2025-02-17", "V003", 9250, 8.0, 3.50, 28.00),
    ("2025-03-03", "V003", 9680, 8.6, 3.55, 30.53),
    ("2025-03-17", "V003", 10090, 8.3, 3.52, 29.22),
    # V004 – van, moderate
    ("2025-01-04", "V004", 32000, 14.0, 3.65, 51.10),
    ("2025-01-18", "V004", 32310, 13.5, 3.70, 49.95),
    ("2025-02-01", "V004", 32640, 14.2, 3.62, 51.40),
    ("2025-02-15", "V004", 32960, 13.8, 3.68, 50.78),
    ("2025-03-01", "V004", 33280, 14.5, 3.72, 53.94),
    ("2025-03-15", "V004", 33590, 13.6, 3.66, 49.78),
    # V005 – old sedan, poor mileage
    ("2025-01-07", "V005", 98000, 14.5, 3.50, 50.75),
    ("2025-01-21", "V005", 98250, 14.8, 3.55, 52.54),
    ("2025-02-04", "V005", 98480, 15.0, 3.48, 52.20),
    ("2025-02-18", "V005", 98720, 14.2, 3.52, 49.98),
    ("2025-03-04", "V005", 98970, 14.6, 3.58, 52.27),
    ("2025-03-18", "V005", 99200, 15.2, 3.54, 53.81),
    # V006 – compact, good mileage
    ("2025-01-08", "V006", 15000, 9.8, 3.40, 33.32),
    ("2025-01-22", "V006", 15380, 9.5, 3.45, 32.78),
    ("2025-02-05", "V006", 15760, 10.0, 3.42, 34.20),
    ("2025-02-19", "V006", 16130, 9.2, 3.48, 32.02),
    ("2025-03-05", "V006", 16520, 9.7, 3.52, 34.14),
    ("2025-03-19", "V006", 16890, 9.4, 3.50, 32.90),
    # V007 – SUV, lower mileage
    ("2025-01-09", "V007", 55000, 17.0, 3.75, 63.75),
    ("2025-01-23", "V007", 55290, 16.5, 3.80, 62.70),
    ("2025-02-06", "V007", 55570, 17.2, 3.78, 65.02),
    ("2025-02-20", "V007", 55860, 16.8, 3.82, 64.18),
    ("2025-03-06", "V007", 56140, 17.5, 3.85, 67.38),
    ("2025-03-20", "V007", 56420, 16.2, 3.79, 61.40),
    # V008 – delivery van, heavy use
    ("2025-01-02", "V008", 72000, 18.0, 3.95, 71.10),
    ("2025-01-16", "V008", 72280, 17.5, 4.00, 70.00),
    ("2025-02-01", "V008", 72550, 18.2, 3.92, 71.34),
    ("2025-02-15", "V008", 72840, 17.8, 3.98, 70.84),
    ("2025-03-01", "V008", 73110, 18.5, 4.05, 74.93),
    ("2025-03-15", "V008", 73400, 17.3, 3.97, 68.68),
    # V009 – electric hybrid, best mileage
    ("2025-01-10", "V009", 5000, 7.5, 3.38, 25.35),
    ("2025-01-24", "V009", 5450, 7.2, 3.42, 24.62),
    ("2025-02-07", "V009", 5890, 7.8, 3.40, 26.52),
    ("2025-02-21", "V009", 6340, 7.0, 3.45, 24.15),
    ("2025-03-07", "V009", 6800, 7.5, 3.50, 26.25),
    ("2025-03-21", "V009", 7240, 7.3, 3.48, 25.40),
    # V010 – pickup, poor mileage
    ("2025-01-11", "V010", 62000, 17.5, 4.10, 71.75),
    ("2025-01-25", "V010", 62230, 18.0, 4.15, 74.70),
    ("2025-02-08", "V010", 62470, 17.2, 4.08, 70.18),
    ("2025-02-22", "V010", 62700, 17.8, 4.12, 73.34),
    ("2025-03-08", "V010", 62940, 28.0, 4.18, 117.04),  # unusually large fill-up (outlier)
    ("2025-03-22", "V010", 63180, 17.0, 4.10, 69.70),
]

HEADERS = ["Date", "Vehicle ID", "Odometer", "Gallons", "Price/Gallon", "Total Cost"]


def create_workbook():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fuel Log"

    # ── Headers ──
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # ── Data rows ──
    for r, row in enumerate(RAW_DATA, 2):
        date_str, vid, odo, gal, ppg, total = row
        ws.cell(row=r, column=1, value=date_str)
        ws.cell(row=r, column=2, value=vid)
        ws.cell(row=r, column=3, value=odo)
        ws.cell(row=r, column=4, value=gal)
        ws.cell(row=r, column=5, value=round(ppg, 2))
        ws.cell(row=r, column=6, value=round(total, 2))

    # ── Column widths ──
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    # Number formats
    for r in range(2, len(RAW_DATA) + 2):
        ws.cell(row=r, column=5).number_format = '$#,##0.00'
        ws.cell(row=r, column=6).number_format = '$#,##0.00'

    ws.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"Saved initial workbook to {OUTPUT_PATH}")


def launch_gui(command, delay_sec=2.0):
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


if __name__ == "__main__":
    create_workbook()
    launch_gui(f'libreoffice --calc "{OUTPUT_PATH}"', delay_sec=2.0)
    print("LibreOffice Calc launched.")
