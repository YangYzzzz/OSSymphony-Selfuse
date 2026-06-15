"""
Initial Setup: Create a spreadsheet with department variance data for number formatting task.
Task ID: calc_lf_086
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_086'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Report'

    # Headers
    ws['A1'] = 'Department'
    ws['B1'] = 'Variance'

    # Style headers
    header_font = Font(bold=True)
    ws['A1'].font = header_font
    ws['B1'].font = header_font

    # Data
    data = [
        ('Sales', 5200),
        ('Marketing', 0),
        ('IT', -1800),
        ('HR', 0),
    ]
    for r, (dept, var) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=2, value=var)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
