"""
Reward Script: Apply custom number format to hide zero values
Task ID: calc_lf_086
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): All four cells B2:B5 have the custom format '#,##0;-#,##0;'
  - Component 2 (0.3): Cell values are preserved (5200, 0, -1800, 0)
           combined with format applied (guards against data corruption)
  - Component 3 (0.3): Non-target cells (A1:A5, B1) retain 'General' format
           combined with at least one B-cell having the custom format
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_086'
EXPECTED_FORMAT = '#,##0;-#,##0;'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Report' sheet must exist
    if 'Report' not in wb.sheetnames:
        print(f"FAIL: 'Report' sheet not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: All four cells B2:B5 have the custom number format (0.4 points)
    # This is the primary task-introduced change.
    try:
        target_cells = ['B2', 'B3', 'B4', 'B5']
        formatted_count = 0
        for coord in target_cells:
            fmt = ws[coord].number_format
            if fmt == EXPECTED_FORMAT:
                formatted_count += 1
                print(f"  PASS: {coord} has format '{fmt}'")
            else:
                print(f"  FAIL: {coord} has format '{fmt}', expected '{EXPECTED_FORMAT}'")

        if formatted_count == 4:
            print(f"PASS: Component 1 — All 4 cells have custom format (0.4 pts)")
            total_score += 0.4
        elif formatted_count > 0:
            partial = round(0.4 * (formatted_count / 4), 2)
            print(f"PARTIAL: Component 1 — {formatted_count}/4 cells formatted ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells have the custom format")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data values preserved AND format applied (0.3 points)
    # Verifies that applying the format didn't corrupt cell values.
    # This component only awards points if at least one cell has the custom format
    # (so it scores 0 on initial_env where format is General).
    try:
        expected_values = {'B2': 5200, 'B3': 0, 'B4': -1800, 'B5': 0}
        values_correct = 0
        for coord, expected in expected_values.items():
            val = ws[coord].value
            if val is not None and abs(float(val) - expected) < 0.01:
                values_correct += 1
            else:
                print(f"  FAIL: {coord} value={val!r}, expected={expected}")

        # Only award points if format is also applied (anchors to task change)
        if values_correct == 4 and formatted_count >= 1:
            print(f"PASS: Component 2 — All values preserved with format applied (0.3 pts)")
            total_score += 0.3
        elif values_correct == 4 and formatted_count == 0:
            print(f"FAIL: Component 2 — Values correct but format not applied (precondition not met)")
        else:
            print(f"FAIL: Component 2 — {values_correct}/4 values correct")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Non-target cells retain General format AND at least one B-cell
    # has the custom format (0.3 points)
    # This ensures the format was applied selectively, not to the whole sheet.
    # Anchored to the task change: only scores if custom format is present on B cells.
    try:
        non_target = ['A1', 'A2', 'A3', 'A4', 'A5', 'B1']
        all_general = True
        for coord in non_target:
            fmt = ws[coord].number_format
            if fmt != 'General':
                print(f"  FAIL: {coord} has format '{fmt}', expected 'General'")
                all_general = False

        if all_general and formatted_count >= 1:
            print(f"PASS: Component 3 — Non-target cells retain General format (0.3 pts)")
            total_score += 0.3
        elif all_general and formatted_count == 0:
            print(f"FAIL: Component 3 — Non-target cells are General but no B-cells have custom format")
        else:
            print(f"FAIL: Component 3 — Some non-target cells have unexpected formats")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state("libreoffice_calc")

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
