"""
Initial Setup: ArXiv cs.CL evolution study — open Chrome and LibreOffice Calc
Task ID: osworld_multi_apps_arxiv_llms_calc_015
Domain: libreoffice_calc (multi-app: Chrome + LibreOffice Calc)

Creates evolution_study.ods with:
  - Sheet1: headers only (no data rows — agent must add 20 rows)
  - Analysis: skeleton structure with month labels and metric columns (no formulas, no chart)
Then opens Chrome pointing at the ArXiv cs.CL listing and opens the file in LibreOffice Calc.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_arxiv_llms_calc_015'
OUTPUT = f'{WORKDIR}/evolution_study.ods'


def launch_gui(command: str, delay_sec: float = 1.5):
    """Launch a GUI app on the VM display without blocking script exit."""
    env = os.environ.copy()
    env['DISPLAY'] = ':0'
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ------------------------------------------------------------------ #
    # Sheet1: Data sheet — headers only, no data rows
    # ------------------------------------------------------------------ #
    ws1 = wb.active
    ws1.title = 'Sheet1'

    headers = ['arXiv ID', 'Title', 'Author Count', 'Abstract Word Count', 'Month']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=False)

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Column widths for readability
    ws1.column_dimensions['A'].width = 20   # arXiv ID
    ws1.column_dimensions['B'].width = 50   # Title
    ws1.column_dimensions['C'].width = 15   # Author Count
    ws1.column_dimensions['D'].width = 22   # Abstract Word Count
    ws1.column_dimensions['E'].width = 12   # Month
    ws1.row_dimensions[1].height = 20

    # ------------------------------------------------------------------ #
    # Analysis sheet: skeleton — month labels and column headers
    # No formulas, no chart. Agent must add those.
    # ------------------------------------------------------------------ #
    ws2 = wb.create_sheet('Analysis')

    # Title row
    ws2['A1'] = 'ArXiv cs.CL Evolution Study — Analysis'
    ws2['A1'].font = Font(bold=True, size=13)

    # Column headers for summary table
    summary_headers = ['Month', 'Avg Author Count', 'Avg Abstract Word Count', '% Change Author Count', '% Change Abstract Word Count']
    sh_font = Font(bold=True, size=11)
    sh_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    sh_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(summary_headers, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = sh_font
        cell.fill = sh_fill
        cell.alignment = sh_align

    # Month label rows (no formulas yet)
    ws2.cell(row=4, column=1, value='January 2024')
    ws2.cell(row=5, column=1, value='April 2024')
    ws2.cell(row=6, column=1, value='% Change')

    # Style month labels
    month_font = Font(bold=False, size=11)
    for r in [4, 5, 6]:
        ws2.cell(row=r, column=1).font = month_font

    # Column widths
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['C'].width = 26
    ws2.column_dimensions['D'].width = 26
    ws2.column_dimensions['E'].width = 30

    # Note row indicating what the agent should do
    ws2['A8'] = 'Note: Add AVERAGE formulas, percentage change formula, and a line chart here.'
    ws2['A8'].font = Font(italic=True, color='FF808080', size=10)

    # ------------------------------------------------------------------ #
    # Save file
    # ------------------------------------------------------------------ #
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # ------------------------------------------------------------------ #
    # GUI startup: open Chrome pointing at ArXiv cs.CL Jan 2024
    # then open the .ods in LibreOffice Calc
    # ------------------------------------------------------------------ #
    launch_gui(
        'google-chrome --new-window "https://arxiv.org/list/cs.CL/2024-01"',
        delay_sec=2.0,
    )
    launch_gui(
        f'libreoffice --calc "{OUTPUT}"',
        delay_sec=2.0,
    )
    print('GUI_READY: Chrome and LibreOffice Calc launched with DISPLAY=:0')


create_initial()
