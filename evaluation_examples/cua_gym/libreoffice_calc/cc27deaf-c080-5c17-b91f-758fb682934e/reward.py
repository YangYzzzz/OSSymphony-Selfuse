"""
Reward Script: ArXiv cs.CL Evolution Study in LibreOffice Calc
Task ID: osworld_multi_apps_arxiv_llms_calc_015
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.35): Sheet1 has 20 data rows (10 per month: 2024-01 and 2024-04)
  - Component 2 (0.35): Analysis sheet has AVERAGE formulas for author count and
                        abstract word count for both January and April
  - Component 3 (0.15): Analysis sheet has percentage change formulas
  - Component 4 (0.15): Analysis sheet has a line chart comparing the two months
Total: 1.0
"""

import os
import shutil

# WORKDIR is the VM path where the task file resides
WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_arxiv_llms_calc_015'
FILE_NAME = 'evolution_study.ods'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The file is .ods in name but xlsx in format (created by openpyxl),
    so we copy it to a .xlsx path before loading.
    """
    import openpyxl

    total_score = 0.0

    # Copy .ods to .xlsx for openpyxl to process (file is actually xlsx-format)
    tmp_path = '/tmp/_reward_evolution_study.xlsx'
    try:
        shutil.copy(file_path, tmp_path)
    except Exception as e:
        print(f"CRITICAL: Cannot copy file {file_path} to tmp: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Load workbook (formula version for formula checks)
    try:
        wb = openpyxl.load_workbook(tmp_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {tmp_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets must exist
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: 'Sheet1' not found in workbook")
        print("REWARD: 0.0")
        return 0.0
    if 'Analysis' not in wb.sheetnames:
        print("CRITICAL: 'Analysis' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws_sheet1 = wb['Sheet1']
    ws_analysis = wb['Analysis']

    # -------------------------------------------------------------------------
    # Component 1: Sheet1 has 20 data rows (10 per month) (0.35 points)
    # This FAILS on initial (Sheet1 has only headers) and PASSES on golden.
    # We count rows with Month values '2024-01' or '2024-04'.
    # -------------------------------------------------------------------------
    try:
        jan_count = 0
        apr_count = 0
        for row in ws_sheet1.iter_rows(min_row=2, max_row=ws_sheet1.max_row):
            month_val = row[4].value  # Column E = index 4
            if month_val is not None:
                month_str = str(month_val).strip()
                if month_str == '2024-01':
                    jan_count += 1
                elif month_str == '2024-04':
                    apr_count += 1

        total_data_rows = jan_count + apr_count
        if total_data_rows >= 20 and jan_count >= 10 and apr_count >= 10:
            print(f"PASS: Component 1 — Sheet1 has {total_data_rows} data rows "
                  f"({jan_count} Jan, {apr_count} Apr) (0.35 pts)")
            total_score += 0.35
        elif total_data_rows > 0:
            # Partial credit proportional to how many rows are present
            partial = round(0.35 * (total_data_rows / 20), 4)
            print(f"PARTIAL: Component 1 — Sheet1 has {total_data_rows}/20 data rows "
                  f"({jan_count} Jan, {apr_count} Apr) ({partial} pts)")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 1 — Sheet1 has no data rows (expected 20)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Analysis sheet has AVERAGE formulas for author count and
    #              abstract word count per month (0.35 points)
    # Checks: rows 4 and 5 (Jan and Apr), columns B and C have AVERAGE formulas
    # referencing Sheet1 data.
    # -------------------------------------------------------------------------
    try:
        avg_checks_passed = 0
        avg_total = 4  # B4, C4, B5, C5

        # B4: avg author count for Jan
        b4 = ws_analysis.cell(row=4, column=2).value
        if b4 and isinstance(b4, str) and 'AVERAGE' in b4.upper() and 'Sheet1' in b4:
            print(f"  PASS: Analysis!B4 has AVERAGE formula for Jan author count: {b4}")
            avg_checks_passed += 1
        else:
            print(f"  FAIL: Analysis!B4 expected AVERAGE(Sheet1!...) formula, found: {b4}")

        # C4: avg abstract word count for Jan
        c4 = ws_analysis.cell(row=4, column=3).value
        if c4 and isinstance(c4, str) and 'AVERAGE' in c4.upper() and 'Sheet1' in c4:
            print(f"  PASS: Analysis!C4 has AVERAGE formula for Jan abstract word count: {c4}")
            avg_checks_passed += 1
        else:
            print(f"  FAIL: Analysis!C4 expected AVERAGE(Sheet1!...) formula, found: {c4}")

        # B5: avg author count for Apr
        b5 = ws_analysis.cell(row=5, column=2).value
        if b5 and isinstance(b5, str) and 'AVERAGE' in b5.upper() and 'Sheet1' in b5:
            print(f"  PASS: Analysis!B5 has AVERAGE formula for Apr author count: {b5}")
            avg_checks_passed += 1
        else:
            print(f"  FAIL: Analysis!B5 expected AVERAGE(Sheet1!...) formula, found: {b5}")

        # C5: avg abstract word count for Apr
        c5 = ws_analysis.cell(row=5, column=3).value
        if c5 and isinstance(c5, str) and 'AVERAGE' in c5.upper() and 'Sheet1' in c5:
            print(f"  PASS: Analysis!C5 has AVERAGE formula for Apr abstract word count: {c5}")
            avg_checks_passed += 1
        else:
            print(f"  FAIL: Analysis!C5 expected AVERAGE(Sheet1!...) formula, found: {c5}")

        comp2_score = round(0.35 * (avg_checks_passed / avg_total), 4)
        if avg_checks_passed == avg_total:
            print(f"PASS: Component 2 — All {avg_total} AVERAGE formulas present (0.35 pts)")
            total_score += comp2_score
        elif avg_checks_passed > 0:
            print(f"PARTIAL: Component 2 — {avg_checks_passed}/{avg_total} AVERAGE formulas "
                  f"present ({comp2_score} pts)")
            total_score += comp2_score
        else:
            print(f"FAIL: Component 2 — No AVERAGE formulas found in Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Analysis sheet has percentage change formulas (0.15 points)
    # The task requires computing percentage change between Jan and Apr for each metric.
    # This should appear in columns D and/or E, or as a separate row.
    # The golden has: D5/E5 = =(B5-B4)/B4*100, =(C5-C4)/C4*100
    # -------------------------------------------------------------------------
    try:
        pct_checks_passed = 0

        # Check row 5 columns D and E for % change formulas
        d5 = ws_analysis.cell(row=5, column=4).value
        e5 = ws_analysis.cell(row=5, column=5).value

        # Also check row 6 (the '% Change' label row) columns D and E
        d6 = ws_analysis.cell(row=6, column=4).value
        e6 = ws_analysis.cell(row=6, column=5).value

        def has_pct_change_formula(val):
            """Check if value looks like a percentage change formula."""
            if val is None:
                return False
            if not isinstance(val, str):
                return False
            val_upper = val.upper().replace(' ', '')
            # Must be a formula and reference B or C columns (from the AVERAGE rows)
            return val.startswith('=') and ('B4' in val or 'B5' in val or 'C4' in val or 'C5' in val)

        # Count percentage change formula occurrences
        pct_formula_cells = []
        for row_idx in range(4, 10):  # rows 4-9
            for col_idx in range(1, 8):  # columns A-G
                cell_val = ws_analysis.cell(row=row_idx, column=col_idx).value
                if has_pct_change_formula(cell_val):
                    coord = ws_analysis.cell(row=row_idx, column=col_idx).coordinate
                    pct_formula_cells.append((coord, cell_val))

        if len(pct_formula_cells) >= 2:
            print(f"PASS: Component 3 — Found {len(pct_formula_cells)} percentage change "
                  f"formula(s): {pct_formula_cells} (0.15 pts)")
            total_score += 0.15
        elif len(pct_formula_cells) == 1:
            print(f"PARTIAL: Component 3 — Found 1/2 percentage change formula: "
                  f"{pct_formula_cells} (0.075 pts)")
            total_score += 0.075
        else:
            # Also check if there's any formula with percentage-related content
            # (e.g., using division by reference cells)
            pct_alt_cells = []
            for row_idx in range(3, 10):
                for col_idx in range(1, 8):
                    cell_val = ws_analysis.cell(row=row_idx, column=col_idx).value
                    if (cell_val and isinstance(cell_val, str)
                            and cell_val.startswith('=')
                            and '100' in cell_val
                            and ('B' in cell_val or 'C' in cell_val)):
                        coord = ws_analysis.cell(row=row_idx, column=col_idx).coordinate
                        pct_alt_cells.append((coord, cell_val))
            if pct_alt_cells:
                print(f"PARTIAL: Component 3 — Found alternative % formula(s): "
                      f"{pct_alt_cells} (0.075 pts)")
                total_score += 0.075
            else:
                print(f"FAIL: Component 3 — No percentage change formulas found in Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Analysis sheet has a line chart (0.15 points)
    # The task requires a line chart comparing Jan vs Apr on both metrics.
    # -------------------------------------------------------------------------
    try:
        charts = ws_analysis._charts
        found_line_chart = False
        chart_series_count = 0

        for chart in charts:
            chart_type = type(chart).__name__
            if 'Line' in chart_type or 'line' in chart_type.lower():
                found_line_chart = (len(chart.series) >= 0)  # derive from actual API
                chart_series_count = len(chart.series)
                break
            # Some charts may appear as BarChart but with line subtype; also check
            # if there's any chart at all

        has_any_chart = len(charts) > 0

        if found_line_chart and chart_series_count >= 2:
            print(f"PASS: Component 4 — Line chart with {chart_series_count} series found "
                  f"in Analysis sheet (0.15 pts)")
            total_score += 0.15
        elif found_line_chart:
            print(f"PARTIAL: Component 4 — Line chart found but only {chart_series_count} "
                  f"series (expected >= 2) (0.1 pts)")
            total_score += 0.1
        elif has_any_chart:
            print(f"PARTIAL: Component 4 — A chart exists but is not a LineChart "
                  f"(type: {type(charts[0]).__name__}) (0.075 pts)")
            total_score += 0.075
        else:
            print(f"FAIL: Component 4 — No chart found in Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Clean up temp file
    try:
        os.remove(tmp_path)
    except Exception:
        pass

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Main entry point
file_path = f'{WORKDIR}/{FILE_NAME}'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
