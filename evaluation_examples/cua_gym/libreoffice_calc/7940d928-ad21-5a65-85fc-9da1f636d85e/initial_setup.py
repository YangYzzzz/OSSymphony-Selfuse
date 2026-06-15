"""
Initial Setup: Create an unprotected 'Test Cases' sheet with 30 test cases.
Task ID: calc_ps_042
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Test Cases'

    # --- Headers ---
    headers = ['TC ID', 'Description', 'Priority', 'Status', 'Assignee']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 20

    # --- 30 Test Cases ---
    test_cases = [
        ['TC-001', 'Verify user login with valid credentials', 'High', 'Passed', 'Sarah Chen'],
        ['TC-002', 'Validate password reset email delivery', 'High', 'Passed', 'Marcus Johnson'],
        ['TC-003', 'Check session timeout after 30 minutes of inactivity', 'Medium', 'Failed', 'Aisha Patel'],
        ['TC-004', 'Test multi-factor authentication via SMS', 'High', 'In Progress', 'David Kim'],
        ['TC-005', 'Verify account lockout after 5 failed attempts', 'Critical', 'Passed', 'Elena Rodriguez'],
        ['TC-006', 'Validate credit card payment processing', 'Critical', 'Passed', 'James O\'Brien'],
        ['TC-007', 'Test shopping cart quantity update functionality', 'Medium', 'Failed', 'Priya Sharma'],
        ['TC-008', 'Check order confirmation email content accuracy', 'Medium', 'Passed', 'Tyler Washington'],
        ['TC-009', 'Verify inventory deduction after purchase', 'High', 'In Progress', 'Sarah Chen'],
        ['TC-010', 'Test coupon code application at checkout', 'Low', 'Not Started', 'Marcus Johnson'],
        ['TC-011', 'Validate search results pagination', 'Medium', 'Passed', 'Aisha Patel'],
        ['TC-012', 'Check product image zoom on hover', 'Low', 'Passed', 'David Kim'],
        ['TC-013', 'Test responsive layout on tablet (768px)', 'High', 'Failed', 'Elena Rodriguez'],
        ['TC-014', 'Verify breadcrumb navigation accuracy', 'Low', 'Passed', 'James O\'Brien'],
        ['TC-015', 'Validate form field input masks for phone numbers', 'Medium', 'In Progress', 'Priya Sharma'],
        ['TC-016', 'Test file upload with 10MB PDF document', 'High', 'Passed', 'Tyler Washington'],
        ['TC-017', 'Check error message display for invalid email format', 'Medium', 'Passed', 'Sarah Chen'],
        ['TC-018', 'Verify API rate limiting at 100 requests per minute', 'High', 'Failed', 'Marcus Johnson'],
        ['TC-019', 'Test data export to CSV format', 'Medium', 'Passed', 'Aisha Patel'],
        ['TC-020', 'Validate user profile photo upload and crop', 'Low', 'Not Started', 'David Kim'],
        ['TC-021', 'Check notification bell badge count accuracy', 'Medium', 'In Progress', 'Elena Rodriguez'],
        ['TC-022', 'Test drag-and-drop file reordering in dashboard', 'Low', 'Not Started', 'James O\'Brien'],
        ['TC-023', 'Verify SSL certificate validation on all endpoints', 'Critical', 'Passed', 'Priya Sharma'],
        ['TC-024', 'Test concurrent user access with 50 sessions', 'High', 'Failed', 'Tyler Washington'],
        ['TC-025', 'Validate date picker localization for EU format', 'Medium', 'Passed', 'Sarah Chen'],
        ['TC-026', 'Check accessibility compliance (WCAG 2.1 AA)', 'High', 'In Progress', 'Marcus Johnson'],
        ['TC-027', 'Test webhook delivery for order status changes', 'Medium', 'Passed', 'Aisha Patel'],
        ['TC-028', 'Verify database backup restoration procedure', 'Critical', 'Passed', 'David Kim'],
        ['TC-029', 'Test dark mode toggle persistence across sessions', 'Low', 'Not Started', 'Elena Rodriguez'],
        ['TC-030', 'Validate bulk user import from CSV with 500 records', 'High', 'Failed', 'James O\'Brien'],
    ]

    for r, row_data in enumerate(test_cases, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Sheet is NOT protected (initial state)
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
