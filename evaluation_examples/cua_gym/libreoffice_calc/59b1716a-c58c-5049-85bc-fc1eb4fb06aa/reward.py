"""
Reward Script: Pivot table with monthly expense totals in chronological order
Task ID: calc_pivot_063
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): Pivot sheet exists with proper structure (headers)
  Component 2 (0.3): Months in chronological order (Jan first, Dec last)
  Component 3 (0.3): Correct monthly SUM of Amount values
  Component 4 (0.2): Grand Total = 230000
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_063'

# Expected month order (chronological) and expected totals
EXPECTED_MONTHS = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December'
]

EXPECTED_TOTALS = {
    'January': 18000,
    'February': 16500,
    'March': 19500,
    'April': 17000,
    'May': 20000,
    'June': 18500,
    'July': 21000,
    'August': 19000,
    'September': 17500,
    'October': 20500,
    'November': 20500,
    'December': 22000,
}

EXPECTED_GRAND_TOTAL = 230000


def find_pivot_sheet(wb):
    """Find the pivot/summary sheet (not MonthlyExp). Returns worksheet or None."""
    for name in wb.sheetnames:
        if name.lower() != 'monthlyexp':
            ws = wb[name]
            # Check if it looks like a pivot table (has month names and numbers)
            # Scan first 20 rows for month names
            for r in range(1, min(ws.max_row + 1, 25)):
                val = ws.cell(row=r, column=1).value
                if val and str(val).strip() in EXPECTED_MONTHS:
                    return ws, name
    return None, None


def find_data_start_row(ws):
    """Find the row where month data starts."""
    for r in range(1, min(ws.max_row + 1, 25)):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip() in EXPECTED_MONTHS:
            return r
    return None


def find_amount_column(ws, header_row):
    """Find which column contains the sum/amount values. Check the row above data start."""
    # Look at header row (row before first month)
    if header_row and header_row > 1:
        for c in range(1, min(ws.max_column + 1, 10)):
            val = ws.cell(row=header_row - 1, column=c).value
            if val and isinstance(val, str):
                lower_val = val.lower()
                if 'amount' in lower_val or 'sum' in lower_val or 'total' in lower_val:
                    return c
    # Default: column 2 (most common pivot layout)
    return 2


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot sheet exists with proper structure (0.2 points)
    try:
        pivot_ws, pivot_name = find_pivot_sheet(wb)
        if pivot_ws is not None:
            data_start = find_data_start_row(pivot_ws)
            if data_start is not None:
                print(f"PASS: Component 1 — Pivot sheet '{pivot_name}' found with data starting at row {data_start} (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 1 — Sheet '{pivot_name}' found but no month data rows detected")
        else:
            print("FAIL: Component 1 — No pivot/summary sheet found (only 'MonthlyExp' exists)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_ws is None:
        # No pivot sheet means nothing else can pass
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    data_start = find_data_start_row(pivot_ws)
    if data_start is None:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    amount_col = find_amount_column(pivot_ws, data_start)

    # Component 2: Months in chronological order (0.3 points)
    try:
        actual_months = []
        for r in range(data_start, data_start + 12):
            val = pivot_ws.cell(row=r, column=1).value
            if val is not None:
                actual_months.append(str(val).strip())

        if len(actual_months) == 12 and actual_months == EXPECTED_MONTHS:
            print(f"PASS: Component 2 — All 12 months in chronological order (0.3 pts)")
            total_score += 0.3
        elif len(actual_months) == 12:
            # Check if they're at least all present
            if set(actual_months) == set(EXPECTED_MONTHS):
                # All months present but wrong order
                # Check if it's alphabetical (common mistake)
                if actual_months == sorted(EXPECTED_MONTHS):
                    print(f"FAIL: Component 2 — Months are in alphabetical order, not chronological")
                else:
                    print(f"FAIL: Component 2 — Months present but not in chronological order: {actual_months}")
            else:
                print(f"FAIL: Component 2 — Missing or extra months: {actual_months}")
        else:
            print(f"FAIL: Component 2 — Expected 12 months, found {len(actual_months)}: {actual_months}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Correct monthly totals (0.3 points)
    try:
        correct_count = 0
        for i, month in enumerate(EXPECTED_MONTHS):
            r = data_start + i
            month_label = pivot_ws.cell(row=r, column=1).value
            amount_val = pivot_ws.cell(row=r, column=amount_col).value

            if month_label and str(month_label).strip() == month:
                expected = EXPECTED_TOTALS[month]
                if amount_val is not None:
                    try:
                        actual = float(amount_val)
                        if abs(actual - expected) < 1.0:
                            correct_count += 1
                        else:
                            print(f"  INFO: {month} amount mismatch: expected {expected}, got {actual}")
                    except (ValueError, TypeError):
                        print(f"  INFO: {month} amount not numeric: {amount_val}")
                else:
                    print(f"  INFO: {month} amount is None")
            else:
                # Month label doesn't match — skip (already penalized in Component 2)
                pass

        if correct_count == 12:
            print(f"PASS: Component 3 — All 12 monthly totals correct (0.3 pts)")
            total_score += 0.3
        elif correct_count >= 6:
            partial = round(0.3 * (correct_count / 12), 2)
            print(f"PARTIAL: Component 3 — {correct_count}/12 monthly totals correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {correct_count}/12 monthly totals correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total = 230000 (0.2 points)
    try:
        gt_label = None
        gt_val = None
        # Search rows after month data for a grand total
        for r in range(data_start + 12, min(pivot_ws.max_row + 1, data_start + 20)):
            label = pivot_ws.cell(row=r, column=1).value
            if label and 'total' in str(label).lower():
                gt_label = str(label).strip()
                gt_val = pivot_ws.cell(row=r, column=amount_col).value
                break

        if gt_label is not None and gt_val is not None:
            try:
                actual_gt = float(gt_val)
                if abs(actual_gt - EXPECTED_GRAND_TOTAL) < 1.0:
                    print(f"PASS: Component 4 — Grand Total = {actual_gt} (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 4 — Grand Total = {actual_gt}, expected {EXPECTED_GRAND_TOTAL}")
            except (ValueError, TypeError):
                print(f"FAIL: Component 4 — Grand Total value not numeric: {gt_val}")
        else:
            print(f"FAIL: Component 4 — No Grand Total row found after month data")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
