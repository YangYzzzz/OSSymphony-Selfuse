"""
Initial Setup: Monthly expense data for pivot table creation task
Task ID: calc_pivot_063
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'MonthlyExp'

    # Headers
    headers = ['ExpID', 'Month', 'Department', 'ExpenseType', 'Amount']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Month order and target totals (summing to 230000)
    months_order = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    monthly_totals = {
        'January': 18000, 'February': 16500, 'March': 19500,
        'April': 17000, 'May': 20000, 'June': 18500,
        'July': 21000, 'August': 19000, 'September': 17500,
        'October': 20500, 'November': 20500, 'December': 22000
    }

    departments = ['Engineering', 'Marketing', 'Finance', 'Operations', 'HR',
                    'Sales', 'Legal', 'IT Support']
    expense_types = ['Travel', 'Office Supplies', 'Software Licenses', 'Training',
                     'Meals & Entertainment', 'Equipment', 'Consulting', 'Utilities',
                     'Advertising', 'Maintenance']

    # Generate 20 rows per month, controlling amounts to hit exact totals
    rows = []
    for month in months_order:
        target = monthly_totals[month]
        # Generate 19 random amounts, then compute the 20th to hit exact target
        amounts = []
        for i in range(19):
            remaining_slots = 20 - len(amounts)
            remaining_budget = target - sum(amounts)
            avg = remaining_budget / remaining_slots
            lo = max(100, int(avg * 0.3))
            hi = min(int(avg * 1.7), remaining_budget - (remaining_slots - 1) * 100)
            if hi < lo:
                hi = lo + 100
            amt = random.randint(lo, hi)
            amounts.append(amt)
        # last amount to hit exact total
        amounts.append(target - sum(amounts))

        random.shuffle(amounts)
        for amt in amounts:
            dept = random.choice(departments)
            exp_type = random.choice(expense_types)
            rows.append((month, dept, exp_type, amt))

    # Shuffle all rows so months are interleaved (realistic unsorted data)
    random.shuffle(rows)

    # Write data rows
    for idx, (month, dept, exp_type, amt) in enumerate(rows, 1):
        r = idx + 1  # row 2 onwards
        ws.cell(row=r, column=1, value=idx)           # ExpID
        ws.cell(row=r, column=2, value=month)          # Month
        ws.cell(row=r, column=3, value=dept)           # Department
        ws.cell(row=r, column=4, value=exp_type)       # ExpenseType
        ws.cell(row=r, column=5, value=amt)            # Amount

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 12

    # Number format for Amount column
    for r in range(2, 242):
        ws.cell(row=r, column=5).number_format = '#,##0'

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
