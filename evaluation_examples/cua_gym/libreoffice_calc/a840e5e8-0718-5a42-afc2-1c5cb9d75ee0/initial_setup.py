"""
Initial Setup: Travel expense report with receipt categorization and per-diem limits
Task ID: calc_gpm_050
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_050'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TravelExp"

    # --- Title row: Merge A1:H1 ---
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Business Travel Expense Report"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FF404040", end_color="FF404040", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 2: Trip info ---
    ws["A2"] = "Employee: John Smith"
    ws["D2"] = "Trip: Client Visit - Chicago"
    ws["G2"] = "Dates: Mar 25-28, 2026"
    for c in ["A2", "D2", "G2"]:
        ws[c].font = Font(size=11)

    # --- Row 4: Headers ---
    headers = ["Date", "Category", "Description", "Amount",
               "Per Diem Limit", "Reimbursable", "Over Limit?", "Receipt?"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Limits lookup table J1:K5 ---
    ws["J1"] = "Category"
    ws["K1"] = "Limit"
    ws["J1"].font = Font(bold=True)
    ws["K1"].font = Font(bold=True)
    limits = [
        ("Meals", 75),
        ("Lodging", 200),
        ("Transport", 150),
        ("Misc", 50),
    ]
    for i, (cat, lim) in enumerate(limits, 2):
        ws.cell(row=i, column=10, value=cat)
        ws.cell(row=i, column=11, value=lim)
        ws.cell(row=i, column=11).number_format = '$#,##0.00'

    # --- 16 expense items rows 5-20 across 4 days ---
    expenses = [
        # Day 1: Mar 25
        ("2026-03-25", "Transport", "Uber to O'Hare Airport", 42.50, "Yes"),
        ("2026-03-25", "Transport", "Flight ORD roundtrip baggage fee", 85.00, "Yes"),
        ("2026-03-25", "Meals", "Lunch at O'Hare Terminal 3", 18.75, "Yes"),
        ("2026-03-25", "Lodging", "Hilton Chicago Downtown - Night 1", 189.00, "Yes"),
        # Day 2: Mar 26
        ("2026-03-26", "Meals", "Breakfast at hotel restaurant", 24.50, "Yes"),
        ("2026-03-26", "Transport", "Taxi to client office", 32.00, "Yes"),
        ("2026-03-26", "Meals", "Working lunch with client team", 67.80, "Yes"),
        ("2026-03-26", "Misc", "Parking validation at client site", 15.00, "No"),
        ("2026-03-26", "Lodging", "Hilton Chicago Downtown - Night 2", 189.00, "Yes"),
        # Day 3: Mar 27
        ("2026-03-27", "Meals", "Breakfast room service", 38.90, "Yes"),
        ("2026-03-27", "Transport", "Rental car for site visits", 165.00, "Yes"),
        ("2026-03-27", "Meals", "Dinner at Gibson's Steakhouse", 92.40, "No"),
        ("2026-03-27", "Misc", "Office supplies for presentation", 47.25, "Yes"),
        ("2026-03-27", "Lodging", "Hilton Chicago Downtown - Night 3", 215.00, "Yes"),
        # Day 4: Mar 28
        ("2026-03-28", "Meals", "Breakfast at Portillo's", 16.30, "Yes"),
        ("2026-03-28", "Transport", "Uber to O'Hare Airport return", 48.50, "Yes"),
    ]

    dollar_fmt = '$#,##0.00'
    for i, (date, cat, desc, amt, receipt) in enumerate(expenses, 5):
        ws.cell(row=i, column=1, value=date)
        ws.cell(row=i, column=2, value=cat)
        ws.cell(row=i, column=3, value=desc)
        cell_d = ws.cell(row=i, column=4, value=amt)
        cell_d.number_format = dollar_fmt
        # Columns E, F left empty (formulas go in golden)
        # Column G left empty (formula goes in golden)
        ws.cell(row=i, column=8, value=receipt)
        # Apply borders to data rows
        for col in range(1, 9):
            ws.cell(row=i, column=col).border = header_border

    # Format E and F columns as currency (even though empty in initial)
    for row in range(5, 21):
        ws.cell(row=row, column=5).number_format = dollar_fmt
        ws.cell(row=row, column=6).number_format = dollar_fmt

    # --- Data Validation: Category dropdown on B5:B20 ---
    dv_cat = DataValidation(
        type="list",
        formula1='"Meals,Lodging,Transport,Misc"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_cat.prompt = "Select expense category"
    dv_cat.promptTitle = "Category"
    dv_cat.add("B5:B20")
    ws.add_data_validation(dv_cat)

    # --- Data Validation: Receipt dropdown on H5:H20 ---
    dv_receipt = DataValidation(
        type="list",
        formula1='"Yes,No"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_receipt.prompt = "Receipt attached?"
    dv_receipt.promptTitle = "Receipt"
    dv_receipt.add("H5:H20")
    ws.add_data_validation(dv_receipt)

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 10

    # Row 1 height for title
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
