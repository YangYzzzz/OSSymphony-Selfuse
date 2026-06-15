"""
Reward Script: Build a travel expense report with receipt categorization, per-diem limits, and approval formatting.
Task ID: calc_gpm_050
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP formulas in E5:E20 (0.25 pts)
  Component 2: MIN formulas in F5:F20 (0.15 pts)
  Component 3: IF formulas in G5:G20 (0.15 pts)
  Component 4: Summary rows 22-23 with formulas (0.20 pts)
  Component 5: Conditional formatting on G and H columns (0.10 pts)
  Component 6: Row 22 double border + D23 red bold (0.15 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_050'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'TravelExp' not in wb.sheetnames:
        print("FAIL: Sheet 'TravelExp' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TravelExp']

    # Component 1: VLOOKUP formulas in E5:E20 (0.25 points)
    # Initial env has E5:E20 all None; golden has VLOOKUP formulas
    try:
        vlookup_count = 0
        for r in range(5, 21):
            val = ws.cell(row=r, column=5).value  # column E
            if val is not None and isinstance(val, str) and 'VLOOKUP' in val.upper():
                vlookup_count += 1
        if vlookup_count >= 14:
            print(f"PASS: Component 1 — {vlookup_count}/16 VLOOKUP formulas in E5:E20 (0.25 pts)")
            total_score += 0.25
        elif vlookup_count >= 8:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {vlookup_count}/16 VLOOKUP formulas in E5:E20 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count}/16 VLOOKUP formulas in E5:E20")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: MIN formulas in F5:F20 (0.15 points)
    # Initial env has F5:F20 all None; golden has MIN(D,E) formulas
    try:
        min_count = 0
        for r in range(5, 21):
            val = ws.cell(row=r, column=6).value  # column F
            if val is not None and isinstance(val, str) and 'MIN' in val.upper():
                min_count += 1
        if min_count >= 14:
            print(f"PASS: Component 2 — {min_count}/16 MIN formulas in F5:F20 (0.15 pts)")
            total_score += 0.15
        elif min_count >= 8:
            partial = 0.08
            print(f"PARTIAL: Component 2 — {min_count}/16 MIN formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {min_count}/16 MIN formulas in F5:F20")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: IF formulas in G5:G20 (0.15 points)
    # Initial env has G5:G20 all None; golden has IF(D>E,"YES","NO") formulas
    try:
        if_count = 0
        for r in range(5, 21):
            val = ws.cell(row=r, column=7).value  # column G
            if val is not None and isinstance(val, str) and 'IF' in val.upper():
                if_count += 1
        if if_count >= 14:
            print(f"PASS: Component 3 — {if_count}/16 IF formulas in G5:G20 (0.15 pts)")
            total_score += 0.15
        elif if_count >= 8:
            partial = 0.08
            print(f"PARTIAL: Component 3 — {if_count}/16 IF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {if_count}/16 IF formulas in G5:G20")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Summary rows 22-23 with formulas (0.20 points)
    # Initial env has rows 22-23 completely empty; golden has labels + SUM/subtraction formulas
    try:
        comp4_score = 0.0

        # Check row 22 label "Total Claimed" in A22
        a22 = ws.cell(row=22, column=1).value
        if a22 is not None and 'total' in str(a22).lower() and 'claim' in str(a22).lower():
            comp4_score += 0.03
            print(f"  PASS: A22 label found: {a22}")
        else:
            print(f"  FAIL: A22 expected 'Total Claimed', found: {a22}")

        # Check D22 has SUM formula
        d22 = ws.cell(row=22, column=4).value
        if d22 is not None and isinstance(d22, str) and 'SUM' in d22.upper():
            comp4_score += 0.05
            print(f"  PASS: D22 SUM formula: {d22}")
        else:
            print(f"  FAIL: D22 expected SUM formula, found: {d22}")

        # Check "Total Reimbursable" label somewhere in row 22 (E22)
        e22 = ws.cell(row=22, column=5).value
        if e22 is not None and 'reimburs' in str(e22).lower():
            comp4_score += 0.02
            print(f"  PASS: E22 label found: {e22}")
        else:
            print(f"  FAIL: E22 expected 'Total Reimbursable', found: {e22}")

        # Check F22 has SUM formula
        f22 = ws.cell(row=22, column=6).value
        if f22 is not None and isinstance(f22, str) and 'SUM' in f22.upper():
            comp4_score += 0.05
            print(f"  PASS: F22 SUM formula: {f22}")
        else:
            print(f"  FAIL: F22 expected SUM formula, found: {f22}")

        # Check row 23 label "Amount Over Limits" in A23
        a23 = ws.cell(row=23, column=1).value
        if a23 is not None and 'over' in str(a23).lower() and 'limit' in str(a23).lower():
            comp4_score += 0.02
            print(f"  PASS: A23 label found: {a23}")
        else:
            print(f"  FAIL: A23 expected 'Amount Over Limits', found: {a23}")

        # Check D23 has subtraction formula (D22-F22)
        d23 = ws.cell(row=23, column=4).value
        if d23 is not None and isinstance(d23, str) and ('D22' in d23.upper() or 'F22' in d23.upper()):
            comp4_score += 0.03
            print(f"  PASS: D23 formula: {d23}")
        else:
            print(f"  FAIL: D23 expected D22-F22 formula, found: {d23}")

        if comp4_score > 0:
            print(f"PASS: Component 4 — Summary rows ({comp4_score:.2f} pts)")
            total_score += comp4_score
        else:
            print(f"FAIL: Component 4 — Summary rows empty")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on G and H columns (0.10 points)
    # Initial env has NO conditional formatting; golden has rules on G5:G20 and H5:H20
    try:
        cf_ranges = []
        for cf in ws.conditional_formatting:
            cf_ranges.append(str(cf))

        has_g_cf = any('G' in r for r in cf_ranges)
        has_h_cf = any('H' in r for r in cf_ranges)

        if has_g_cf and has_h_cf:
            print(f"PASS: Component 5 — Conditional formatting on G and H columns (0.10 pts)")
            total_score += 0.10
        elif has_g_cf or has_h_cf:
            print(f"PARTIAL: Component 5 — Conditional formatting on {'G' if has_g_cf else 'H'} only (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting found (ranges: {cf_ranges})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Row 22 double border + D23 red bold formatting (0.15 points)
    # Initial env: no borders on row 22, D23 not bold, no red font
    try:
        comp6_score = 0.0

        # Check double border on row 22 (at least top or bottom should be "double")
        border_count = 0
        for col_idx in range(1, 9):
            c = ws.cell(row=22, column=col_idx)
            if c.border.top.style == 'double' or c.border.bottom.style == 'double':
                border_count += 1
        if border_count >= 4:
            comp6_score += 0.07
            print(f"  PASS: Row 22 double border found on {border_count}/8 cells")
        else:
            print(f"  FAIL: Row 22 double border found on only {border_count}/8 cells")

        # Check D23 bold
        d23_cell = ws.cell(row=23, column=4)
        if d23_cell.font.bold:
            comp6_score += 0.04
            print(f"  PASS: D23 is bold")
        else:
            print(f"  FAIL: D23 is not bold")

        # Check D23 red font color
        try:
            font_color = d23_cell.font.color.rgb
            if font_color is not None and 'FF0000' in str(font_color).upper():
                comp6_score += 0.04
                print(f"  PASS: D23 has red font color: {font_color}")
            else:
                print(f"  FAIL: D23 font color is not red: {font_color}")
        except Exception:
            print(f"  FAIL: D23 font color not readable")

        if comp6_score > 0:
            print(f"PASS: Component 6 — Formatting ({comp6_score:.2f} pts)")
            total_score += comp6_score
        else:
            print(f"FAIL: Component 6 — No formatting applied")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
