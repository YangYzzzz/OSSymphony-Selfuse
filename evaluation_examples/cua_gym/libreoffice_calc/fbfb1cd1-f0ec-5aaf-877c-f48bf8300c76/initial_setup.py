"""
Initial Setup: HR Merit Raise Calculation
Task ID: calc_hr_merit_raise_calc_026
Domain: libreoffice_calc

Creates initial spreadsheet with employee data for merit review.
Columns E (Raise %), F (Raise Amount), G (New Salary) are empty — to be filled by agent.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_hr_merit_raise_calc_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Merit Review'

    # Row 1: Headers
    headers = ['Emp ID', 'Name', 'Current Salary', 'Rating', 'Raise %', 'Raise Amount', 'New Salary']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 62 employees with realistic data (rows 2-63)
    # Mix of Exceptional, Meets Expectations, Below Expectations ratings
    employees = [
        ('E001', 'Sarah Chen',         82500, 'Exceptional'),
        ('E002', 'Marcus Johnson',     71200, 'Meets Expectations'),
        ('E003', 'Priya Patel',        68900, 'Below Expectations'),
        ('E004', 'Derek Williams',     91500, 'Exceptional'),
        ('E005', 'Elena Vasquez',      74300, 'Meets Expectations'),
        ('E006', 'James Okafor',       63800, 'Meets Expectations'),
        ('E007', 'Anita Krishnaswamy', 88700, 'Exceptional'),
        ('E008', 'Robert Nguyen',      76400, 'Below Expectations'),
        ('E009', 'Linda Park',         59200, 'Meets Expectations'),
        ('E010', 'Thomas Brennan',     105000,'Exceptional'),
        ('E011', 'Fatima Al-Rashid',   67500, 'Meets Expectations'),
        ('E012', 'Carlos Reyes',       79800, 'Exceptional'),
        ('E013', 'Mei Huang',          54300, 'Below Expectations'),
        ('E014', 'Brian Sullivan',     93200, 'Meets Expectations'),
        ('E015', 'Nkechi Obi',         70600, 'Exceptional'),
        ('E016', 'David Kowalski',     61400, 'Meets Expectations'),
        ('E017', 'Rachel Goldstein',   84100, 'Exceptional'),
        ('E018', 'Ahmed Hassan',       72900, 'Below Expectations'),
        ('E019', 'Yuki Tanaka',        66700, 'Meets Expectations'),
        ('E020', 'Michael Torres',     97300, 'Exceptional'),
        ('E021', 'Chloe Dubois',       58900, 'Meets Expectations'),
        ('E022', 'Samuel Adeyemi',     80200, 'Exceptional'),
        ('E023', 'Ingrid Larsson',     75600, 'Below Expectations'),
        ('E024', 'Patrick O\'Brien',   69400, 'Meets Expectations'),
        ('E025', 'Zara Mohammed',      87200, 'Exceptional'),
        ('E026', 'Andrew Kim',         62700, 'Meets Expectations'),
        ('E027', 'Diana Fernandez',    94800, 'Exceptional'),
        ('E028', 'Jerome Washington',  73500, 'Below Expectations'),
        ('E029', 'Leila Nazari',       55800, 'Meets Expectations'),
        ('E030', 'Gregory Petrov',     88400, 'Meets Expectations'),
        ('E031', 'Amara Diallo',       67900, 'Exceptional'),
        ('E032', 'Sean Murphy',        78600, 'Below Expectations'),
        ('E033', 'Hana Yamamoto',      61200, 'Meets Expectations'),
        ('E034', 'Victor Okonkwo',     99500, 'Exceptional'),
        ('E035', 'Sophia Russo',       76800, 'Meets Expectations'),
        ('E036', 'Kevin Chandra',      65400, 'Exceptional'),
        ('E037', 'Maria Santos',       83700, 'Meets Expectations'),
        ('E038', 'Daniel Johansson',   70100, 'Below Expectations'),
        ('E039', 'Aisha Kamara',       58500, 'Meets Expectations'),
        ('E040', 'Christopher Lee',    91700, 'Exceptional'),
        ('E041', 'Natasha Ivanova',    74900, 'Meets Expectations'),
        ('E042', 'Kwame Mensah',       63100, 'Below Expectations'),
        ('E043', 'Isabella Moretti',   86300, 'Exceptional'),
        ('E044', 'Noah Andersen',      77500, 'Meets Expectations'),
        ('E045', 'Riya Mehta',         69800, 'Exceptional'),
        ('E046', 'Jason Brooks',       57600, 'Below Expectations'),
        ('E047', 'Valentina Cruz',     92400, 'Meets Expectations'),
        ('E048', 'Oliver Schmidt',     80900, 'Exceptional'),
        ('E049', 'Tunde Afolabi',      64700, 'Meets Expectations'),
        ('E050', 'Megan O\'Sullivan',  72300, 'Exceptional'),
        ('E051', 'Kenji Watanabe',     85600, 'Meets Expectations'),
        ('E052', 'Alicia Rivera',      60400, 'Below Expectations'),
        ('E053', 'Stefan Novak',       78200, 'Meets Expectations'),
        ('E054', 'Blessing Eze',       66900, 'Exceptional'),
        ('E055', 'Luca Bianchi',       95700, 'Meets Expectations'),
        ('E056', 'Yasmin Khalil',      71800, 'Below Expectations'),
        ('E057', 'Connor MacPherson',  83200, 'Exceptional'),
        ('E058', 'Adaeze Nwosu',       59700, 'Meets Expectations'),
        ('E059', 'Henrik Lindqvist',   89100, 'Exceptional'),
        ('E060', 'Catalina Gomez',     68300, 'Meets Expectations'),
        ('E061', 'Rashid Al-Farsi',    76100, 'Below Expectations'),
        ('E062', 'Nnenna Okafor',      81400, 'Exceptional'),
    ]

    for r, (emp_id, name, salary, rating) in enumerate(employees, 2):
        ws.cell(row=r, column=1, value=emp_id)
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=3, value=salary)
        ws.cell(row=r, column=4, value=rating)
        # Columns E (5), F (6), G (7) are intentionally left empty

    # Row 64 is empty (reserved for totals — to be filled by agent)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Merit Review')
    print(f'  Rows 2-63: {len(employees)} employees')
    print(f'  Columns E, F, G: empty (task targets)')


create_initial()
