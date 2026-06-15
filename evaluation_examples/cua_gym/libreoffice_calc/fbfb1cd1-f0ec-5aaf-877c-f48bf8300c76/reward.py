"""
Reward Script: HR Merit Raise Calculator
Task ID: calc_hr_merit_raise_calc_026
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 — IFS raise percentage formulas in E2:E63          0.30 pts
  Component 2 — Raise amount formulas in F2:F63 (=C*E)           0.25 pts
  Component 3 — New salary formulas in G2:G63 (=C+F)             0.25 pts
  Component 4 — Total merit budget row (A64 label + F64 SUM)     0.20 pts
  Total:                                                           1.00 pts

All four components FAIL on initial file (columns E/F/G are empty, row 64 is blank).
All four components PASS on golden file.
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_hr_merit_raise_calc_026'
SHEET_NAME = 'Merit Review'
DATA_ROWS = range(2, 64)  # rows 2..63 inclusive (62 employees)


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # -------------------------------------------------------------------
    # Component 1: IFS raise percentage formulas in E2:E63 (0.30 points)
    # Each row must have =IFS(D{r}="Exceptional",0.05,D{r}="Meets Expectations",0.03,
    #                         D{r}="Below Expectations",0)
    # and number_format must be a percentage format (contains '%')
    # -------------------------------------------------------------------
    try:
        e_formula_correct = 0
        e_format_correct = 0
        e_total = 62  # rows 2..63

        for r in DATA_ROWS:
            cell = ws.cell(row=r, column=5)
            val = cell.value
            fmt = cell.number_format if cell.number_format else ''

            # Check formula (case-insensitive, whitespace-normalized)
            expected = (
                f'=IFS(D{r}="Exceptional",0.05,'
                f'D{r}="Meets Expectations",0.03,'
                f'D{r}="Below Expectations",0)'
            )
            if isinstance(val, str):
                norm_val = val.strip().upper().replace(' ', '')
                norm_exp = expected.strip().upper().replace(' ', '')
                if norm_val == norm_exp:
                    e_formula_correct += 1

            if '%' in fmt:
                e_format_correct += 1

        formula_ratio = e_formula_correct / e_total
        format_ratio = e_format_correct / e_total

        if formula_ratio == 1.0 and format_ratio == 1.0:
            print(f"PASS: Component 1 — All {e_total} IFS formulas in E2:E63 correct "
                  f"with percentage format (0.30 pts)")
            total_score += 0.30
        elif formula_ratio >= 0.9 and format_ratio >= 0.9:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {e_formula_correct}/{e_total} correct formulas, "
                  f"{e_format_correct}/{e_total} correct formats (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {e_formula_correct}/{e_total} IFS formulas "
                  f"correct, {e_format_correct}/{e_total} percentage formats. "
                  f"Expected =IFS(D2=\"Exceptional\",0.05,...) in E2:E63")
    except Exception as e:
        print(f"ERROR: Component 1 (IFS formulas E2:E63) — {e}")

    # -------------------------------------------------------------------
    # Component 2: Raise amount formulas in F2:F63 (0.25 points)
    # Each row must have =C{r}*E{r} and number_format '$#,##0.00'
    # -------------------------------------------------------------------
    try:
        f_formula_correct = 0
        f_format_correct = 0
        f_total = 62

        for r in DATA_ROWS:
            cell = ws.cell(row=r, column=6)
            val = cell.value
            fmt = cell.number_format if cell.number_format else ''

            expected = f'=C{r}*E{r}'
            if isinstance(val, str):
                norm_val = val.strip().upper().replace(' ', '')
                norm_exp = expected.strip().upper().replace(' ', '')
                if norm_val == norm_exp:
                    f_formula_correct += 1

            if '$' in fmt and '#' in fmt:
                f_format_correct += 1

        formula_ratio = f_formula_correct / f_total
        format_ratio = f_format_correct / f_total

        if formula_ratio == 1.0 and format_ratio == 1.0:
            print(f"PASS: Component 2 — All {f_total} raise amount formulas in F2:F63 "
                  f"correct with currency format (0.25 pts)")
            total_score += 0.25
        elif formula_ratio >= 0.9 and format_ratio >= 0.9:
            partial = 0.12
            print(f"PARTIAL: Component 2 — {f_formula_correct}/{f_total} correct formulas, "
                  f"{f_format_correct}/{f_total} correct formats (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {f_formula_correct}/{f_total} formulas correct, "
                  f"{f_format_correct}/{f_total} currency formats. "
                  f"Expected =C2*E2 (and variants) in F2:F63 with $#,##0.00 format")
    except Exception as e:
        print(f"ERROR: Component 2 (raise amount formulas F2:F63) — {e}")

    # -------------------------------------------------------------------
    # Component 3: New salary formulas in G2:G63 (0.25 points)
    # Each row must have =C{r}+F{r} and number_format '$#,##0.00'
    # -------------------------------------------------------------------
    try:
        g_formula_correct = 0
        g_format_correct = 0
        g_total = 62

        for r in DATA_ROWS:
            cell = ws.cell(row=r, column=7)
            val = cell.value
            fmt = cell.number_format if cell.number_format else ''

            expected = f'=C{r}+F{r}'
            if isinstance(val, str):
                norm_val = val.strip().upper().replace(' ', '')
                norm_exp = expected.strip().upper().replace(' ', '')
                if norm_val == norm_exp:
                    g_formula_correct += 1

            if '$' in fmt and '#' in fmt:
                g_format_correct += 1

        formula_ratio = g_formula_correct / g_total
        format_ratio = g_format_correct / g_total

        if formula_ratio == 1.0 and format_ratio == 1.0:
            print(f"PASS: Component 3 — All {g_total} new salary formulas in G2:G63 "
                  f"correct with currency format (0.25 pts)")
            total_score += 0.25
        elif formula_ratio >= 0.9 and format_ratio >= 0.9:
            partial = 0.12
            print(f"PARTIAL: Component 3 — {g_formula_correct}/{g_total} correct formulas, "
                  f"{g_format_correct}/{g_total} correct formats (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {g_formula_correct}/{g_total} formulas correct, "
                  f"{g_format_correct}/{g_total} currency formats. "
                  f"Expected =C2+F2 (and variants) in G2:G63 with $#,##0.00 format")
    except Exception as e:
        print(f"ERROR: Component 3 (new salary formulas G2:G63) — {e}")

    # -------------------------------------------------------------------
    # Component 4: Total merit budget row (0.20 points)
    # A64 must be 'Total Merit Budget' (label)
    # F64 must be =SUM(F2:F63) (formula) with $#,##0.00 format
    # -------------------------------------------------------------------
    try:
        a64 = ws.cell(row=64, column=1).value
        f64 = ws.cell(row=64, column=6).value
        f64_fmt = ws.cell(row=64, column=6).number_format or ''

        label_ok = isinstance(a64, str) and 'total merit budget' in a64.lower()
        formula_ok = (
            isinstance(f64, str)
            and f64.strip().upper().replace(' ', '') == '=SUM(F2:F63)'
        )
        format_ok = '$' in f64_fmt and '#' in f64_fmt

        if label_ok and formula_ok and format_ok:
            print(f"PASS: Component 4 — A64='Total Merit Budget', "
                  f"F64='=SUM(F2:F63)' with currency format (0.20 pts)")
            total_score += 0.20
        elif label_ok and formula_ok:
            partial = 0.15
            print(f"PARTIAL: Component 4 — Label and SUM formula present but "
                  f"F64 format is '{f64_fmt}' not currency (partial {partial} pts)")
            total_score += partial
        elif formula_ok:
            partial = 0.10
            print(f"PARTIAL: Component 4 — F64 SUM formula present but A64 label "
                  f"is '{a64}' not 'Total Merit Budget' (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — A64='{a64}' (expected 'Total Merit Budget'), "
                  f"F64='{f64}' (expected '=SUM(F2:F63)'), "
                  f"F64 format='{f64_fmt}' (expected currency with $)")
    except Exception as e:
        print(f"ERROR: Component 4 (total merit budget row 64) — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
