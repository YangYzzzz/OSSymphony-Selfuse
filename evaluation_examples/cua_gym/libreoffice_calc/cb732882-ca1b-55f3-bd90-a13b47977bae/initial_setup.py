"""
Initial Setup: Create a spreadsheet with price history data (no chart)
Task ID: calc_chart_line_step_075
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_line_step_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: PriceHistory ---
    ws = wb.active
    ws.title = 'PriceHistory'

    # Headers
    ws['A1'] = 'Date'
    ws['B1'] = 'Unit Price ($)'

    # Style headers
    for cell in [ws['A1'], ws['B1']]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows as specified in context
    # Date strings and prices per task context
    data = [
        ('01-Jan', 29.99),
        ('15-Feb', 29.99),
        ('01-Mar', 34.99),
        ('15-Apr', 34.99),
        ('01-Jun', 39.99),
        ('15-Jul', 39.99),
        ('01-Sep', 44.99),
        ('30-Nov', 44.99),
    ]

    for row_idx, (date_val, price_val) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=date_val)
        ws.cell(row=row_idx, column=2, value=price_val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 18

    # NOTE: No charts exist in the initial file - the task is to CREATE a chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('  Sheet: PriceHistory')
    print('  Rows: 1 header + 8 data rows (rows 1-9)')
    print('  Columns: Date (A), Unit Price ($) (B)')
    print('  No charts present (task is to add a step-line chart)')


create_initial()
