"""
Initial Setup: Create procurement spreadsheet with vendor list for dynamic validation task
Task ID: calc_gcv_094
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # --- Headers in row 1 (columns A-E) ---
    headers = ['PO Number', 'Item', 'Quantity', 'Cost', 'Vendor']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- 19 purchase orders (rows 2-20) ---
    po_data = [
        ['PO-2025-0401', 'Industrial Bearings (50mm)', 120, 4560.00],
        ['PO-2025-0402', 'Hydraulic Hose Assembly', 45, 2835.00],
        ['PO-2025-0403', 'Stainless Steel Flanges', 200, 7800.00],
        ['PO-2025-0404', 'Electrical Conduit (3/4")', 500, 3250.00],
        ['PO-2025-0405', 'Safety Goggles (ANSI Z87)', 300, 1950.00],
        ['PO-2025-0406', 'Welding Wire ER70S-6', 80, 5120.00],
        ['PO-2025-0407', 'PVC Pipe Schedule 40', 350, 2275.00],
        ['PO-2025-0408', 'Carbide End Mill Set', 15, 3375.00],
        ['PO-2025-0409', 'Pneumatic Cylinder (100mm)', 60, 8400.00],
        ['PO-2025-0410', 'Heat Shrink Tubing Kit', 100, 1200.00],
        ['PO-2025-0411', 'Torque Wrench (25-250 ft-lb)', 10, 2890.00],
        ['PO-2025-0412', 'Silicone Sealant (Clear)', 240, 1680.00],
        ['PO-2025-0413', 'Copper Busbar (1/4" x 2")', 30, 4350.00],
        ['PO-2025-0414', 'Vibration Dampening Pads', 150, 2250.00],
        ['PO-2025-0415', 'Thread Sealant Tape (PTFE)', 1000, 1500.00],
        ['PO-2025-0416', 'Precision Linear Rails', 8, 6400.00],
        ['PO-2025-0417', 'Abrasive Cutting Discs', 500, 3750.00],
        ['PO-2025-0418', 'Thermal Grease Compound', 50, 975.00],
        ['PO-2025-0419', 'Spring Pin Assortment Kit', 25, 1125.00],
    ]

    for r, row_data in enumerate(po_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:  # Cost column
                cell.number_format = '$#,##0.00'
        # Column E (Vendor) is intentionally left EMPTY
        ws.cell(row=r, column=5).border = thin_border

    # --- Vendor list in column H (H1:H12) ---
    vendors = [
        'Grainger Industrial',
        'McMaster-Carr',
        'Fastenal Company',
        'MSC Industrial Direct',
        'W.W. Grainger',
        'Applied Industrial Tech',
        'Motion Industries',
        'Kaman Distribution',
        'RS Components',
        'Newark Electronics',
        'Allied Electronics',
        'Digi-Key Corporation',
    ]

    for i, vendor in enumerate(vendors, 1):
        ws.cell(row=i, column=8, value=vendor)

    # --- Column widths ---
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 24
    ws.column_dimensions['H'].width = 24

    # --- No named ranges, no data validation ---

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
