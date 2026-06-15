"""
Initial Setup: Reading Tracker - articles.ods with 5 article URLs
Task ID: osworld_multi_apps_multi_simple_011
Domain: libreoffice_calc (multi-app: Calc + Chrome)
"""

import os
import shlex
import subprocess
import time

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_multi_simple_011'
READING_DIR = '/home/user/reading'
OUTPUT = f'{READING_DIR}/articles.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Create the reading directory
    os.makedirs(READING_DIR, exist_ok=True)

    # Create the articles.ods file using Python with odfpy
    # First try odfpy, fall back to xlsx+convert approach
    try:
        import subprocess as sp
        # Install odfpy if needed
        sp.run(['pip3', 'install', 'odfpy', '--quiet'], capture_output=True)
        from odf.opendocument import OpenDocumentSpreadsheet
        from odf.table import Table, TableRow, TableCell
        from odf.text import P
        from odf.style import Style, TableCellProperties, TextProperties
        from odf import namespaces
        import odf.table

        doc = OpenDocumentSpreadsheet()

        # Define a text style for the header
        header_style = Style(name="HeaderStyle", family="table-cell")
        header_style.addElement(TextProperties(fontweight="bold"))
        doc.styles.addElement(header_style)

        table = Table(name="Articles")

        # Header row
        headers = ['Article_ID', 'Title', 'URL', 'Notes', 'Word_Count']
        header_row = TableRow()
        for h in headers:
            cell = TableCell(stylename="HeaderStyle", valuetype="string")
            cell.addElement(P(text=h))
            header_row.addElement(cell)
        table.addElement(header_row)

        # 5 article rows — realistic content, Notes and Word_Count EMPTY
        articles = [
            ('ART001', 'The Future of Renewable Energy', 'https://www.scientificamerican.com/article/renewable-energy-future/'),
            ('ART002', 'Machine Learning in Healthcare: A Primer', 'https://hbr.org/2018/05/machine-learning-in-healthcare'),
            ('ART003', 'Understanding Climate Change: Key Facts', 'https://www.nationalgeographic.com/environment/article/climate-change-overview'),
            ('ART004', 'The Psychology of Procrastination', 'https://www.psychologytoday.com/us/basics/procrastination'),
            ('ART005', 'Urban Planning for Sustainable Cities', 'https://www.theguardian.com/cities/2019/nov/14/urban-planning-sustainable-cities'),
        ]

        for art_id, title, url in articles:
            row = TableRow()
            for val in [art_id, title, url]:
                cell = TableCell(valuetype="string")
                cell.addElement(P(text=val))
                row.addElement(cell)
            # Notes: empty
            empty_cell = TableCell()
            row.addElement(empty_cell)
            # Word_Count: empty
            empty_cell2 = TableCell()
            row.addElement(empty_cell2)
            table.addElement(row)

        doc.spreadsheet.addElement(table)
        doc.save(OUTPUT)
        print(f'Initial file created with odfpy: {OUTPUT}')

    except Exception as e:
        print(f'odfpy failed ({e}), falling back to xlsx+convert approach')
        # Fallback: create xlsx, then convert to ods using libreoffice headless
        import openpyxl
        tmp_xlsx = f'{READING_DIR}/articles_tmp.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Articles'

        headers = ['Article_ID', 'Title', 'URL', 'Notes', 'Word_Count']
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)

        articles = [
            ('ART001', 'The Future of Renewable Energy', 'https://www.scientificamerican.com/article/renewable-energy-future/', None, None),
            ('ART002', 'Machine Learning in Healthcare: A Primer', 'https://hbr.org/2018/05/machine-learning-in-healthcare', None, None),
            ('ART003', 'Understanding Climate Change: Key Facts', 'https://www.nationalgeographic.com/environment/article/climate-change-overview', None, None),
            ('ART004', 'The Psychology of Procrastination', 'https://www.psychologytoday.com/us/basics/procrastination', None, None),
            ('ART005', 'Urban Planning for Sustainable Cities', 'https://www.theguardian.com/cities/2019/nov/14/urban-planning-sustainable-cities', None, None),
        ]
        for r, row_data in enumerate(articles, 2):
            for c, val in enumerate(row_data, 1):
                if val is not None:
                    ws.cell(row=r, column=c, value=val)

        wb.save(tmp_xlsx)

        # Convert to ods
        env_convert = os.environ.copy()
        env_convert["DISPLAY"] = ":0"
        result = subprocess.run(
            ['soffice', '--headless', '--convert-to', 'ods', '--outdir', READING_DIR, tmp_xlsx],
            env=env_convert,
            capture_output=True,
            text=True,
            timeout=60
        )
        print(f'Conversion stdout: {result.stdout}')
        print(f'Conversion stderr: {result.stderr}')

        # Rename if needed
        converted = f'{READING_DIR}/articles_tmp.ods'
        if os.path.exists(converted):
            os.rename(converted, OUTPUT)
            os.remove(tmp_xlsx)
            print(f'Converted to ODS: {OUTPUT}')
        else:
            # If conversion failed, use xlsx as fallback with .ods extension attempt
            print(f'Conversion output not found, using xlsx as fallback')
            # Just save the xlsx at the ods path as a last resort
            wb.save(OUTPUT.replace('.ods', '.xlsx'))
            print(f'Fallback: created {OUTPUT.replace(".ods", ".xlsx")}')

    # Verify file was created
    if os.path.exists(OUTPUT):
        size = os.path.getsize(OUTPUT)
        print(f'File verified: {OUTPUT} ({size} bytes)')
    else:
        print(f'WARNING: File not found at {OUTPUT}')

    # GUI-ready startup: open the articles.ods with LibreOffice Calc, and also open Chrome
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    launch_gui('google-chrome', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc + Chrome with DISPLAY=:0')


create_initial()
