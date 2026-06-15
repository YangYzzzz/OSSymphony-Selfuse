"""
Initial Setup: sales.xlsx with quarterly revenue data on Desktop
Task ID: osworld_multi_apps_calc_vscode_001
Domain: libreoffice_calc (multi-app: VSCode also required)

Creates:
  - /home/user/Desktop/sales.xlsx  — quarterly revenue spreadsheet
  - Opens it in LibreOffice Calc
  - Opens VSCode (for the agent to write the Python script)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_calc_vscode_001'
OUTPUT = f'{WORKDIR}/sales.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(WORKDIR, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Sheet: Sales ---
    ws = wb.active
    ws.title = 'Sales'

    # Header row
    ws.cell(row=1, column=1, value='Quarter')
    ws.cell(row=1, column=2, value='Revenue')

    # Style header row: bold, blue background
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', size=12)
    header_align = Alignment(horizontal='center', vertical='center')
    for col in range(1, 3):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    # Data rows — realistic quarterly revenue (some blank cells intentionally)
    # Blanks simulate missing/unreported quarters
    data = [
        ('Q1 2023', 125430.50),
        ('Q2 2023', 138920.75),
        ('Q3 2023', 142650.00),
        ('Q4 2023', None),       # blank — unreported quarter
        ('Q1 2024', 151230.25),
        ('Q2 2024', 163840.50),
        ('Q3 2024', None),       # blank — unreported quarter
        ('Q4 2024', 178920.00),
        ('Q1 2025', 183450.75),
        ('Q2 2025', 195680.25),
        ('Q3 2025', 204310.00),
        ('Q4 2025', 217890.50),
    ]

    for row_idx, (quarter, revenue) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=quarter)
        if revenue is not None:
            cell = ws.cell(row=row_idx, column=2, value=revenue)
            cell.number_format = '#,##0.00'
        # blank cells are left empty (None)

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with sales.xlsx
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)

    # Also open VSCode (agent will need it to write the Python script)
    launch_gui('code /home/user', delay_sec=2.0)

    print('GUI_READY: launched LibreOffice Calc and VSCode with DISPLAY=:0')


create_initial()
