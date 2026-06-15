"""
Initial Setup: ACL/ML Award-winning Papers Tracker
Task ID: osworld_multi_apps_acl_awards_calc_014
Domain: libreoffice_calc

Creates global_awards.ods with:
  - Sheet1: headers only (Conference, Year, Title, First Author, Institution, Country)
  - Country Analysis: headers only (Country, Count)
Opens Chrome and LibreOffice Calc for the agent to begin.
"""

import os
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_acl_awards_calc_014'
OUTPUT = f'{WORKDIR}/global_awards.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet1: Award Papers Data ---
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Column headers only — agent must fill in all data rows
    headers = ['Conference', 'Year', 'Title', 'First Author', 'Institution', 'Country']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Set reasonable column widths
    ws1.column_dimensions['A'].width = 14  # Conference
    ws1.column_dimensions['B'].width = 8   # Year
    ws1.column_dimensions['C'].width = 60  # Title
    ws1.column_dimensions['D'].width = 22  # First Author
    ws1.column_dimensions['E'].width = 40  # Institution
    ws1.column_dimensions['F'].width = 20  # Country

    # --- Country Analysis sheet: headers only ---
    ws2 = wb.create_sheet('Country Analysis')
    ws2.cell(row=1, column=1, value='Country')
    ws2.cell(row=1, column=2, value='Count')
    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open Chrome and LibreOffice Calc
    # 1. Open Chrome (agent will use it to look up award papers)
    launch_gui('google-chrome --no-first-run --no-default-browser-check', delay_sec=2.0)

    # 2. Open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=3.0)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
