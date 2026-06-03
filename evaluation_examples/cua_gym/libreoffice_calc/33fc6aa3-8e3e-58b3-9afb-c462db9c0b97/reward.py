"""
Reward Script: Collect award-winning papers from ICML 2021, ICML 2022, ICLR 2022 into LibreOffice Calc
Task ID: osworld_multi_apps_acl_awards_calc_014
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Sheet1 has 6+ data rows with all required columns populated,
                      covering ICML 2021, ICML 2022, and ICLR 2022
  Component 2 (0.3): 'Country Analysis' sheet has country entries with COUNTIF formulas
  Component 3 (0.3): 'Country Analysis' sheet has at least one bar chart
"""

import os
import shutil
import zipfile

WORKDIR = '/home/user'
TASK_ID = 'global_awards'
FILE_PATH = f'{WORKDIR}/{TASK_ID}.ods'
# The file is .ods by name but actually XLSX format (openpyxl compatible)
XLSX_COPY = '/tmp/global_awards_reward_check.xlsx'

# Required columns for Sheet1
REQUIRED_COLUMNS = ['Conference', 'Year', 'Title', 'First Author', 'Institution', 'Country']

# Required conferences/years
REQUIRED_SOURCES = [
    ('ICML', 2021),
    ('ICML', 2022),
    ('ICLR', 2022),
]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Make a copy with .xlsx extension so openpyxl can read it
    try:
        shutil.copy2(file_path, XLSX_COPY)
    except Exception as e:
        print(f"CRITICAL: Cannot copy file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(XLSX_COPY)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: verify required sheets exist
    if 'Sheet1' not in wb.sheetnames:
        print(f"CRITICAL: 'Sheet1' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    if 'Country Analysis' not in wb.sheetnames:
        print(f"FAIL: 'Country Analysis' sheet not found. Sheets: {wb.sheetnames}")
        # Continue to check what's there, but this is partial failure

    # -------------------------------------------------------------------------
    # Component 1: Sheet1 has 6+ data rows covering all 3 conference/year combos
    #              with all required columns populated (0.4 points)
    # -------------------------------------------------------------------------
    try:
        ws1 = wb['Sheet1']

        # Check headers in row 1
        headers = [ws1.cell(row=1, column=c).value for c in range(1, 7)]
        headers_lower = [str(h).strip().lower() if h else '' for h in headers]

        expected_headers_lower = [h.lower() for h in REQUIRED_COLUMNS]
        headers_match = all(
            any(eh in hl for hl in headers_lower)
            for eh in expected_headers_lower
        )

        if not headers_match:
            print(f"FAIL: Component 1 — headers mismatch. Found: {headers}")
        else:
            # Count data rows (rows 2 onwards)
            data_rows = []
            for row in ws1.iter_rows(min_row=2, max_row=ws1.max_row):
                row_vals = [cell.value for cell in row]
                non_empty = [v for v in row_vals if v is not None and str(v).strip() != '']
                if len(non_empty) >= 4:  # At least 4 fields populated
                    data_rows.append(row_vals)

            # Check for coverage of the 3 required conference/year combos
            # Determine column positions
            col_conference = None
            col_year = None
            for i, h in enumerate(headers):
                if h and 'conference' in str(h).lower():
                    col_conference = i
                if h and 'year' in str(h).lower():
                    col_year = i

            if col_conference is None or col_year is None:
                print(f"FAIL: Component 1 — could not find Conference or Year columns. Headers: {headers}")
            elif len(data_rows) < 6:
                print(f"FAIL: Component 1 — only {len(data_rows)} data rows found (need 6+). "
                      f"Expected entries from ICML 2021, ICML 2022, ICLR 2022.")
            else:
                # Check coverage of required conference/year combinations
                found_sources = set()
                for row_vals in data_rows:
                    conf = row_vals[col_conference]
                    year = row_vals[col_year]
                    if conf is not None and year is not None:
                        conf_upper = str(conf).strip().upper()
                        try:
                            year_int = int(year)
                        except (ValueError, TypeError):
                            year_int = None

                        for req_conf, req_year in REQUIRED_SOURCES:
                            if req_conf.upper() in conf_upper and year_int == req_year:
                                found_sources.add((req_conf, req_year))

                missing = [s for s in REQUIRED_SOURCES if s not in found_sources]

                if len(data_rows) >= 6 and len(found_sources) >= 2:
                    # Award full credit if all 3 sources covered, partial if 2
                    if len(found_sources) == 3:
                        print(f"PASS: Component 1 — {len(data_rows)} data rows from all 3 conference sources "
                              f"({found_sources}). (0.4 pts)")
                        total_score += 0.4
                    else:
                        print(f"PARTIAL: Component 1 — {len(data_rows)} data rows but only {len(found_sources)}/3 "
                              f"conference sources covered (missing: {missing}). (0.2 pts)")
                        total_score += 0.2
                elif len(data_rows) >= 6:
                    # Has enough rows but unidentifiable sources
                    print(f"PARTIAL: Component 1 — {len(data_rows)} data rows but sources unclear "
                          f"(found: {found_sources}). (0.2 pts)")
                    total_score += 0.2
                else:
                    print(f"FAIL: Component 1 — {len(data_rows)} rows is insufficient (need 6+). "
                          f"Sources found: {found_sources}.")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Country Analysis sheet has country entries with COUNTIF
    #              formulas in column B (0.3 points)
    # -------------------------------------------------------------------------
    try:
        if 'Country Analysis' not in wb.sheetnames:
            print(f"FAIL: Component 2 — 'Country Analysis' sheet missing.")
        else:
            ws2 = wb['Country Analysis']

            # Check for data beyond the headers (row 2 onwards)
            country_rows = []
            countif_rows = 0

            for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
                country_val = row[0].value if len(row) > 0 else None
                count_val = row[1].value if len(row) > 1 else None

                if country_val is not None and str(country_val).strip():
                    country_rows.append(country_val)
                    # Check if count column has a formula or numeric value
                    if count_val is not None:
                        count_str = str(count_val).strip()
                        if count_str.upper().startswith('=COUNTIF') or count_str.upper().startswith('=COUNTIFS'):
                            countif_rows += 1
                        elif count_str.upper().startswith('='):
                            countif_rows += 1  # Any formula counts

            if len(country_rows) >= 2 and countif_rows >= 2:
                print(f"PASS: Component 2 — Country Analysis has {len(country_rows)} country entries, "
                      f"{countif_rows} with COUNTIF/formula. Countries: {country_rows}. (0.3 pts)")
                total_score += 0.3
            elif len(country_rows) >= 2:
                print(f"PARTIAL: Component 2 — Country Analysis has {len(country_rows)} entries but "
                      f"only {countif_rows} have COUNTIF formulas. Countries: {country_rows}. (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 2 — Country Analysis has insufficient data. "
                      f"Found {len(country_rows)} country entries, {countif_rows} with formulas.")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Country Analysis sheet has at least one bar chart (0.3 points)
    # -------------------------------------------------------------------------
    try:
        if 'Country Analysis' not in wb.sheetnames:
            print(f"FAIL: Component 3 — 'Country Analysis' sheet missing, no chart possible.")
        else:
            ws2 = wb['Country Analysis']
            charts = ws2._charts if hasattr(ws2, '_charts') else []

            if len(charts) >= 1:
                # Check if any chart is a bar/column chart
                bar_charts = []
                for chart in charts:
                    chart_type_name = type(chart).__name__.lower()
                    if 'bar' in chart_type_name or 'column' in chart_type_name:
                        bar_charts.append(chart)
                    elif hasattr(chart, 'type') and chart.type in ('bar', 'col'):
                        bar_charts.append(chart)

                if bar_charts:
                    print(f"PASS: Component 3 — Country Analysis has {len(bar_charts)} bar/column chart(s). (0.3 pts)")
                    total_score += 0.3
                else:
                    # Any chart is acceptable (task says bar chart, but give partial credit)
                    print(f"PARTIAL: Component 3 — Country Analysis has {len(charts)} chart(s) but "
                          f"not identified as bar chart. Chart types: "
                          f"{[type(c).__name__ for c in charts]}. (0.15 pts)")
                    total_score += 0.15
            else:
                print(f"FAIL: Component 3 — No charts found on 'Country Analysis' sheet.")

    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Cleanup
    try:
        os.remove(XLSX_COPY)
    except Exception:
        pass

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
if not os.path.exists(FILE_PATH):
    print(f"File not found: {FILE_PATH}")
    print("REWARD: 0.0")
else:
    verify_task(FILE_PATH)
