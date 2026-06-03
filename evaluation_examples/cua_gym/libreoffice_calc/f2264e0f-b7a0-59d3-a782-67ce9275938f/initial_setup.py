"""
Initial Setup: Create a spreadsheet with 15 months of revenue data for forecast charting.
Task ID: calc_gcp_079
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_079'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Forecast'

    # --- Headers ---
    headers = ['Month', 'Actual', 'Forecast', 'Bridge']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Month labels (A2:A16) ---
    months = [
        'Jan 2024', 'Feb 2024', 'Mar 2024', 'Apr 2024',
        'May 2024', 'Jun 2024', 'Jul 2024', 'Aug 2024',
        'Sep 2024', 'Oct 2024', 'Nov 2024', 'Dec 2024',
        'Jan 2025', 'Feb 2025', 'Mar 2025'
    ]

    # --- Actual revenue values for Jan-Dec 2024 (rows 2-13) ---
    actual_values = [
        42500, 45800, 51200, 48900,
        55300, 62100, 58700, 67400,
        71200, 75800, 82300, 89100
    ]

    # --- Forecast values for Jan-Mar 2025 (rows 14-16) ---
    forecast_values = [92000, 88000, 95000]

    # Write data
    money_fmt = '$#,##0'

    for i, month in enumerate(months):
        row = i + 2
        ws.cell(row=row, column=1, value=month)

        if i < 12:
            # Actual data for Jan-Dec 2024
            c = ws.cell(row=row, column=2, value=actual_values[i])
            c.number_format = money_fmt
            # Column C (Forecast) blank for actual months
            # Column D (Bridge) blank for most months
        else:
            # Column B (Actual) blank for forecast months
            # Forecast values for Jan-Mar 2025
            c = ws.cell(row=row, column=3, value=forecast_values[i - 12])
            c.number_format = money_fmt

    # Bridge column: repeat last actual value at month 12 (Dec 2024, row 13)
    # This allows the forecast line to connect from the last actual data point
    bridge_cell = ws.cell(row=13, column=4, value=89100)
    bridge_cell.number_format = money_fmt

    # Also put bridge value at first forecast month (Jan 2025, row 14)
    # so the forecast line starts from where actual ends
    bridge_cell2 = ws.cell(row=14, column=4, value=89100)
    bridge_cell2.number_format = money_fmt

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14

    # NO charts in initial state - the task is to create the chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
