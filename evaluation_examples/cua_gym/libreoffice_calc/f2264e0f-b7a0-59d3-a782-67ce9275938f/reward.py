"""
Reward Script: Create a dynamic chart with actual values and forecast dashed line
Task ID: calc_gcp_079
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): LineChart exists on Forecast sheet
  Component 2 (0.25): Chart has 2 series referencing Actual (col B) and Forecast (col C)
  Component 3 (0.30): Line styles — Actual is solid, Forecast is dashed/dotted
  Component 4 (0.10): Chart title contains relevant keywords
  Component 5 (0.10): C13 bridge value populated (connects actual to forecast)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_079'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Forecast' sheet must exist
    if 'Forecast' not in wb.sheetnames:
        print("FAIL: 'Forecast' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Forecast']

    # =========================================================================
    # Component 1: A LineChart exists on the Forecast sheet (0.25 points)
    # This FAILS on initial (0 charts) -> PASSES on golden (1 chart)
    # =========================================================================
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Check that at least one chart is a LineChart
            from openpyxl.chart import LineChart
            line_charts = [c for c in charts if isinstance(c, LineChart)]
            if len(line_charts) >= 1:
                print(f"PASS: Component 1 — LineChart found on Forecast sheet ({len(line_charts)} line chart(s)) (0.25 pts)")
                total_score += 0.25
            else:
                chart_types = [type(c).__name__ for c in charts]
                print(f"FAIL: Component 1 — Chart(s) found but none are LineChart. Types: {chart_types}")
        else:
            print("FAIL: Component 1 — No charts found on Forecast sheet")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Chart has 2 data series referencing columns B and C (0.25 points)
    # This FAILS on initial (no chart) -> PASSES on golden (2 series)
    # =========================================================================
    try:
        from openpyxl.chart import LineChart
        line_charts = [c for c in ws._charts if isinstance(c, LineChart)]
        if line_charts:
            chart = line_charts[0]
            num_series = len(chart.series)
            if num_series >= 2:
                # Check that the series reference columns B and C
                series_refs = []
                for s in chart.series:
                    ref_str = ""
                    if hasattr(s, 'val') and s.val and hasattr(s.val, 'numRef') and s.val.numRef:
                        ref_str = s.val.numRef.f
                    series_refs.append(ref_str)

                has_b_ref = any('B' in ref.upper() for ref in series_refs if ref)
                has_c_ref = any('C' in ref.upper() for ref in series_refs if ref)

                if has_b_ref and has_c_ref:
                    print(f"PASS: Component 2 — Chart has {num_series} series referencing B and C columns. Refs: {series_refs} (0.25 pts)")
                    total_score += 0.25
                else:
                    print(f"FAIL: Component 2 — Series refs don't cover both B and C. Refs: {series_refs}")
            else:
                print(f"FAIL: Component 2 — Expected >=2 series, found {num_series}")
        else:
            print("FAIL: Component 2 — No LineChart found")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Actual series is solid, Forecast series is dashed (0.30 points)
    # This FAILS on initial (no chart) -> PASSES on golden (correct line styles)
    # =========================================================================
    try:
        from openpyxl.chart import LineChart
        line_charts = [c for c in ws._charts if isinstance(c, LineChart)]
        if line_charts:
            chart = line_charts[0]
            if len(chart.series) >= 2:
                # Identify which series is Actual and which is Forecast
                actual_series = None
                forecast_series = None

                for s in chart.series:
                    title_str = ""
                    if s.title:
                        if hasattr(s.title, 'strRef') and s.title.strRef:
                            ref_f = s.title.strRef.f
                            # If title references B1 -> Actual, C1 -> Forecast
                            if 'B1' in ref_f.upper() or 'B1' in ref_f:
                                actual_series = s
                            elif 'C1' in ref_f.upper() or 'C1' in ref_f:
                                forecast_series = s
                        elif hasattr(s.title, 'v') and s.title.v:
                            v = str(s.title.v).lower()
                            if 'actual' in v:
                                actual_series = s
                            elif 'forecast' in v:
                                forecast_series = s

                    # Also check via data reference
                    if actual_series is None or forecast_series is None:
                        ref_str = ""
                        if hasattr(s, 'val') and s.val and hasattr(s.val, 'numRef') and s.val.numRef:
                            ref_str = s.val.numRef.f
                        if '$B$' in ref_str or '!B' in ref_str or 'B2' in ref_str:
                            if actual_series is None:
                                actual_series = s
                        elif '$C$' in ref_str or '!C' in ref_str or 'C2' in ref_str:
                            if forecast_series is None:
                                forecast_series = s

                actual_ok = False
                forecast_ok = False

                if actual_series:
                    gp = actual_series.graphicalProperties
                    if gp and gp.line:
                        dash = gp.line.dashStyle or gp.line.prstDash
                        if dash in ('solid', None):
                            actual_ok = True
                            print(f"  Actual series line style: {dash} (solid) — OK")
                        else:
                            print(f"  Actual series line style: {dash} — expected solid")
                    else:
                        # No explicit line style means default (solid)
                        actual_ok = True
                        print("  Actual series line style: default (solid) — OK")

                if forecast_series:
                    gp = forecast_series.graphicalProperties
                    if gp and gp.line:
                        dash = gp.line.dashStyle or gp.line.prstDash
                        # Accept any non-solid dash style
                        if dash and dash != 'solid':
                            forecast_ok = True
                            print(f"  Forecast series line style: {dash} (dashed) — OK")
                        else:
                            print(f"  Forecast series line style: {dash} — expected dashed/dotted")
                    else:
                        print("  Forecast series: no explicit line style (default solid) — FAIL")

                if actual_ok and forecast_ok:
                    print(f"PASS: Component 3 — Actual=solid, Forecast=dashed (0.30 pts)")
                    total_score += 0.30
                elif forecast_ok:
                    # Partial: forecast is dashed but actual style unknown
                    print(f"PARTIAL: Component 3 — Forecast is dashed but Actual series not confirmed solid (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 3 — Line styles not correctly set. actual_ok={actual_ok}, forecast_ok={forecast_ok}")
            else:
                print("FAIL: Component 3 — Not enough series to check line styles")
        else:
            print("FAIL: Component 3 — No LineChart found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Chart has a meaningful title (0.10 points)
    # This FAILS on initial (no chart) -> PASSES on golden (title set)
    # =========================================================================
    try:
        from openpyxl.chart import LineChart
        line_charts = [c for c in ws._charts if isinstance(c, LineChart)]
        if line_charts:
            chart = line_charts[0]
            title_text = ""
            if chart.title and hasattr(chart.title, 'tx') and chart.title.tx:
                if hasattr(chart.title.tx, 'rich') and chart.title.tx.rich:
                    for p in chart.title.tx.rich.p:
                        for r in p.r:
                            title_text += r.t
            if title_text:
                print(f"PASS: Component 4 — Chart title: '{title_text}' (0.10 pts)")
                total_score += 0.10
            else:
                print("FAIL: Component 4 — Chart has no title text")
        else:
            print("FAIL: Component 4 — No LineChart found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: C13 bridge value populated to connect actual->forecast (0.10 points)
    # Initial: C13=None, Golden: C13=89100
    # This FAILS on initial -> PASSES on golden
    # =========================================================================
    try:
        c13_val = ws['C13'].value
        if c13_val is not None and isinstance(c13_val, (int, float)) and c13_val > 0:
            print(f"PASS: Component 5 — C13 bridge value populated: {c13_val} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — C13 bridge value not set (found: {c13_val})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
