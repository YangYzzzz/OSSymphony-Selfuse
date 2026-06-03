"""
Reward Script: Create a formatted product comparison matrix with feature checkmarks and weighted scoring.
Task ID: calc_gpm_053
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Title merge + formatting (A1:G1 merged, bold 14pt white on purple)
  Component 2 (0.15) - Header row 3 formatting (bold, purple fill, white text)
  Component 3 (0.20) - Weighted Total row 15 (SUMPRODUCT formulas, bold 14pt, 0.00 format)
  Component 4 (0.10) - Rank row 16 (RANK formulas)
  Component 5 (0.10) - Data validation on C4:E13 (whole number 1-5)
  Component 6 (0.15) - Conditional formatting on C4:E13 (5 color rules)
  Component 7 (0.10) - Borders and column widths
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_053'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Compare' sheet must exist
    if 'Compare' not in wb.sheetnames:
        print("CRITICAL: 'Compare' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Compare']

    # Component 1: Title merge + formatting (0.20 points)
    # Golden: A1:G1 merged, value='Software Vendor Comparison Matrix', bold 14pt, white text on purple fill
    # Initial: no merge, no formatting
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_title_merge = any('A1' in r and 'G1' in r for r in merged_ranges)

        a1 = ws['A1']
        has_title_value = a1.value is not None and 'Software Vendor Comparison Matrix' in str(a1.value)
        has_title_bold = a1.font.bold is True
        has_title_size = a1.font.size is not None and a1.font.size >= 14

        # Check purple fill (4B0082 = dark purple)
        try:
            fill_rgb = str(a1.fill.fgColor.rgb).upper()
            has_purple_fill = ('4B0082' in fill_rgb)
        except Exception:
            has_purple_fill = False

        comp1_checks = [has_title_merge, has_title_bold, has_title_size, has_purple_fill]
        comp1_pass = sum(comp1_checks)

        if comp1_pass >= 3:
            print(f"PASS: Component 1 — Title merged & formatted (merge={has_title_merge}, bold={has_title_bold}, size14={has_title_size}, purpleFill={has_purple_fill}) (0.20 pts)")
            total_score += 0.20
        elif comp1_pass >= 2:
            print(f"PARTIAL: Component 1 — Title partially formatted ({comp1_pass}/4 checks) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Title not properly formatted (merge={has_title_merge}, bold={has_title_bold}, size14={has_title_size}, purpleFill={has_purple_fill})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Header row 3 formatting (0.15 points)
    # Golden: A3:F3 bold, purple fill (FF4B0082), white font text
    # Initial: no bold, no fill
    try:
        header_bold_count = 0
        header_fill_count = 0
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            cell = ws[f'{col_letter}3']
            if cell.font.bold is True:
                header_bold_count += 1
            try:
                fill_rgb = cell.fill.fgColor.rgb
                if fill_rgb and '4B0082' in str(fill_rgb).upper():
                    header_fill_count += 1
            except Exception:
                pass

        if header_bold_count >= 5 and header_fill_count >= 5:
            print(f"PASS: Component 2 — Headers formatted (bold={header_bold_count}/6, purpleFill={header_fill_count}/6) (0.15 pts)")
            total_score += 0.15
        elif header_bold_count >= 4:
            print(f"PARTIAL: Component 2 — Headers partially formatted (bold={header_bold_count}/6, fill={header_fill_count}/6) (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2 — Headers not formatted (bold={header_bold_count}/6, fill={header_fill_count}/6)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Weighted Total row 15 with SUMPRODUCT formulas (0.20 points)
    # Golden: A15='Weighted Total' bold, C15:E15 have SUMPRODUCT formulas, bold 14pt, format 0.00
    # Initial: row 15 is empty
    try:
        a15 = ws['A15']
        has_label = a15.value is not None and 'Weighted Total' in str(a15.value)

        formula_count = 0
        format_count = 0
        bold_count = 0
        for col_letter in ['C', 'D', 'E']:
            cell = ws[f'{col_letter}15']
            val = cell.value
            if val is not None and isinstance(val, str) and 'SUMPRODUCT' in val.upper():
                formula_count += 1
            if cell.number_format == '0.00':
                format_count += 1
            if cell.font.bold is True:
                bold_count += 1

        if has_label and formula_count == 3 and format_count >= 2 and bold_count >= 2:
            print(f"PASS: Component 3 — Weighted Total row complete (label={has_label}, formulas={formula_count}/3, fmt={format_count}/3, bold={bold_count}/3) (0.20 pts)")
            total_score += 0.20
        elif has_label and formula_count >= 2:
            print(f"PARTIAL: Component 3 — Weighted Total partially complete (formulas={formula_count}/3) (0.10 pts)")
            total_score += 0.10
        elif formula_count >= 1:
            print(f"PARTIAL: Component 3 — At least one SUMPRODUCT formula found (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No Weighted Total row (label={has_label}, formulas={formula_count})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Rank row 16 (0.10 points)
    # Golden: A16='Rank', C16:E16 have RANK formulas
    # Initial: row 16 is empty
    try:
        a16 = ws['A16']
        has_rank_label = a16.value is not None and 'Rank' in str(a16.value)

        rank_formula_count = 0
        for col_letter in ['C', 'D', 'E']:
            cell = ws[f'{col_letter}16']
            val = cell.value
            if val is not None and isinstance(val, str) and 'RANK' in val.upper():
                rank_formula_count += 1

        if has_rank_label and rank_formula_count >= 2:
            print(f"PASS: Component 4 — Rank row present (label={has_rank_label}, rankFormulas={rank_formula_count}/3) (0.10 pts)")
            total_score += 0.10
        elif rank_formula_count >= 1:
            print(f"PARTIAL: Component 4 — Some rank formulas ({rank_formula_count}/3) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No Rank row (label={has_rank_label}, formulas={rank_formula_count})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Data validation on C4:E13 (0.10 points)
    # Golden: whole number 1-5 with prompt 'Rate 1-5'
    # Initial: no data validations
    try:
        dv_list = ws.data_validations.dataValidation
        found_dv = any(
            dv.type == 'whole' and ('C4' in str(dv.sqref) or 'C4:E13' in str(dv.sqref) or 'C4:E11' in str(dv.sqref))
            for dv in dv_list
        )

        if found_dv:
            print(f"PASS: Component 5 — Data validation found (type=whole, covers score cells) (0.10 pts)")
            total_score += 0.10
        else:
            # Check for any data validation at all
            if len(dv_list) > 0:
                print(f"PARTIAL: Component 5 — Some data validation exists but not matching expected (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 5 — No data validations found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on C4:E13 with color scale (0.15 points)
    # Golden: 5 rules for values 1-5 with different fill colors, plus winner highlight on C15:E15
    # Initial: no conditional formatting
    try:
        cf_list = list(ws.conditional_formatting)
        score_range_cf = False
        total_range_cf = False

        for cf in cf_list:
            cf_range = str(cf)
            if 'C4' in cf_range and 'E13' in cf_range:
                # Check for multiple color rules (should be 5 rules for values 1-5)
                score_range_cf = (len(cf.rules) >= 4)
            if 'C15' in cf_range and 'E15' in cf_range:
                total_range_cf = (len(cf.rules) >= 1)

        if score_range_cf and total_range_cf:
            print(f"PASS: Component 6 — Conditional formatting on scores and totals (0.15 pts)")
            total_score += 0.15
        elif score_range_cf:
            print(f"PARTIAL: Component 6 — Conditional formatting on scores only (0.10 pts)")
            total_score += 0.10
        elif len(cf_list) > 0:
            print(f"PARTIAL: Component 6 — Some conditional formatting exists (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Borders and column widths (0.10 points)
    # Golden: thick borders on outside, adjusted column widths (A=20, B=8, C:E=12, F=25)
    # Initial: no borders, all widths=13
    try:
        # Check for thick borders on matrix edge
        left_border_ok = (ws['A3'].border.left.style == 'thick')
        right_border_ok = (ws['F3'].border.right.style == 'thick')

        # Check column widths changed from default 13
        col_a_width = ws.column_dimensions['A'].width
        col_f_width = ws.column_dimensions['F'].width
        widths_adjusted = (col_a_width is not None and col_a_width > 15) or (col_f_width is not None and col_f_width > 15)

        border_checks = sum([left_border_ok, right_border_ok, widths_adjusted])
        if border_checks >= 2:
            print(f"PASS: Component 7 — Borders and widths (leftThick={left_border_ok}, rightThick={right_border_ok}, widthsAdj={widths_adjusted}) (0.10 pts)")
            total_score += 0.10
        elif border_checks >= 1:
            print(f"PARTIAL: Component 7 — Some border/width formatting ({border_checks}/3 checks) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — No borders or width adjustments (leftThick={left_border_ok}, rightThick={right_border_ok}, widthsAdj={widths_adjusted})")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
