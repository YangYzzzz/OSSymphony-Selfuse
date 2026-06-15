"""
Initial Setup: Create a product comparison matrix with raw data (no formatting, no formulas)
Task ID: calc_gpm_053
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_053'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Compare"

    # Row 1: Title text only (no merge, no formatting - that's the task)
    ws["A1"] = "Software Vendor Comparison Matrix"

    # Row 3: Headers (plain text, no formatting)
    headers = ["Feature", "Weight", "Vendor A", "Vendor B", "Vendor C", "Notes"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Rows 4-13: Feature data with weights (as decimals) and scores (1-5)
    features = [
        ["Pricing", 0.20, 4, 3, 5, "Vendor C offers best value for enterprise tier"],
        ["Ease of Use", 0.15, 5, 4, 3, "Vendor A has intuitive onboarding flow"],
        ["Integration", 0.15, 3, 5, 4, "Vendor B supports 200+ native integrations"],
        ["Security", 0.20, 4, 4, 5, "Vendor C has SOC2 Type II and ISO 27001"],
        ["Support", 0.10, 3, 5, 4, "Vendor B provides 24/7 dedicated support"],
        ["Scalability", 0.10, 4, 3, 5, "Vendor C handles 10M+ concurrent users"],
        ["Customization", 0.05, 5, 3, 4, "Vendor A offers full API and SDK access"],
        ["Mobile App", 0.05, 3, 4, 2, "Vendor B mobile app rated 4.7 on App Store"],
    ]

    for r, row_data in enumerate(features, 4):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            # Format weight column as percentage display
            if c == 2:
                cell.number_format = '0%'

    # Rows 15-16: EMPTY - the task is to add weighted totals and ranks here

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
