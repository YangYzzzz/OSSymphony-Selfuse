"""
Reward Script: Contract management date formulas task
Task ID: calc_gen_dateformulas_062
Domain: libreoffice_calc
Scoring:
  - Component 1: E2:E81 days remaining formula =MAX(0, D-TODAY())        (0.25 pts)
  - Component 2: F2:F81 renewal window formula =IF(...<=60, OPEN/CLOSED) (0.25 pts)
  - Component 3: G2:G81 expiry quarter formula using MONTH/YEAR          (0.15 pts)
  - Component 4: H2:H81 business days formula =NETWORKDAYS(C,D)          (0.15 pts)
  - Component 5: Conditional formatting (red<=30, orange<=60, yellow<=90) (0.10 pts)
  - Component 6: Data sorted ascending by Days Remaining / Expiry Date   (0.10 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_dateformulas_062'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'Contracts' must exist
    if 'Contracts' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Contracts' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contracts']

    # Precondition: must have 81 rows (header + 80 data rows) and 8 columns
    if ws.max_row < 81 or ws.max_column < 8:
        print(f"CRITICAL: Sheet dimensions too small: {ws.max_row}x{ws.max_column}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: E2:E81 — Days Remaining formula =MAX(0, D?-TODAY()) (0.25 points)
    # Checks: all 80 rows have a formula containing MAX, 0, and TODAY()
    try:
        e_correct = 0
        e_issues = []
        for row in range(2, 82):
            val = ws.cell(row=row, column=5).value
            if (val is not None and isinstance(val, str)
                    and 'MAX(0' in val.upper()
                    and 'TODAY()' in val.upper()):
                e_correct += 1
            else:
                e_issues.append((row, val))

        if e_correct == 80:
            print(f"PASS: Component 1 — All 80 rows in E have MAX(0,D?-TODAY()) formula (0.25 pts)")
            total_score += 0.25
        elif e_correct >= 60:
            partial = round(0.25 * e_correct / 80, 4)
            print(f"PARTIAL: Component 1 — {e_correct}/80 rows in E have correct formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {e_correct}/80 E rows have correct formula. First issues: {e_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2:F81 — Renewal Window formula =IF(D?-TODAY()<=60,"OPEN","CLOSED") (0.25 points)
    # Checks: all 80 rows have IF formula with <=60, OPEN, CLOSED
    try:
        f_correct = 0
        f_issues = []
        for row in range(2, 82):
            val = ws.cell(row=row, column=6).value
            if (val is not None and isinstance(val, str)
                    and 'IF(' in val.upper()
                    and '60' in val
                    and 'OPEN' in val.upper()
                    and 'CLOSED' in val.upper()):
                f_correct += 1
            else:
                f_issues.append((row, val))

        if f_correct == 80:
            print(f"PASS: Component 2 — All 80 rows in F have IF(<=60,OPEN/CLOSED) formula (0.25 pts)")
            total_score += 0.25
        elif f_correct >= 60:
            partial = round(0.25 * f_correct / 80, 4)
            print(f"PARTIAL: Component 2 — {f_correct}/80 rows in F have correct formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {f_correct}/80 F rows have correct formula. First issues: {f_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2:G81 — Expiry Quarter formula using MONTH() and YEAR() (0.15 points)
    # Checks: all 80 rows have a formula containing MONTH( and YEAR(
    try:
        g_correct = 0
        g_issues = []
        for row in range(2, 82):
            val = ws.cell(row=row, column=7).value
            if (val is not None and isinstance(val, str)
                    and 'MONTH(' in val.upper()
                    and 'YEAR(' in val.upper()):
                g_correct += 1
            else:
                g_issues.append((row, val))

        if g_correct == 80:
            print(f"PASS: Component 3 — All 80 rows in G have quarter label formula (MONTH/YEAR) (0.15 pts)")
            total_score += 0.15
        elif g_correct >= 60:
            partial = round(0.15 * g_correct / 80, 4)
            print(f"PARTIAL: Component 3 — {g_correct}/80 rows in G have correct formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {g_correct}/80 G rows have correct formula. First issues: {g_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: H2:H81 — Business Days formula =NETWORKDAYS(C?,D?) (0.15 points)
    # Checks: all 80 rows have a NETWORKDAYS formula
    try:
        h_correct = 0
        h_issues = []
        for row in range(2, 82):
            val = ws.cell(row=row, column=8).value
            if (val is not None and isinstance(val, str)
                    and 'NETWORKDAYS(' in val.upper()):
                h_correct += 1
            else:
                h_issues.append((row, val))

        if h_correct == 80:
            print(f"PASS: Component 4 — All 80 rows in H have NETWORKDAYS formula (0.15 pts)")
            total_score += 0.15
        elif h_correct >= 60:
            partial = round(0.15 * h_correct / 80, 4)
            print(f"PARTIAL: Component 4 — {h_correct}/80 rows in H have correct formula ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {h_correct}/80 H rows have correct formula. First issues: {h_issues[:3]}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting — 3 rules for <=30 (red), <=60 (orange), <=90 (yellow) (0.10 points)
    # Checks: at least 3 CF rules referencing the E column thresholds 30, 60, 90
    try:
        cf_rules = ws.conditional_formatting
        all_rules = []
        for cf_range, rule_list in cf_rules._cf_rules.items():
            for rule in rule_list:
                all_rules.append(rule)

        # Collect formula-based CF rules that reference E column thresholds
        found_30 = sum(
            1 for rule in all_rules
            if any('<=30' in str(f).upper() or '<= 30' in str(f).upper()
                   for f in (getattr(rule, 'formula', None) or []))
        )
        found_60 = sum(
            1 for rule in all_rules
            if any('<=60' in str(f).upper() or '<= 60' in str(f).upper()
                   for f in (getattr(rule, 'formula', None) or []))
        )
        found_90 = sum(
            1 for rule in all_rules
            if any('<=90' in str(f).upper() or '<= 90' in str(f).upper()
                   for f in (getattr(rule, 'formula', None) or []))
        )
        has_30 = found_30 >= 1
        has_60 = found_60 >= 1
        has_90 = found_90 >= 1

        cf_count = sum([has_30, has_60, has_90])
        if cf_count == 3:
            print(f"PASS: Component 5 — All 3 CF rules present (<=30 red, <=60 orange, <=90 yellow) (0.10 pts)")
            total_score += 0.10
        elif cf_count >= 2:
            print(f"PARTIAL: Component 5 — {cf_count}/3 CF thresholds found (<=30={has_30}, <=60={has_60}, <=90={has_90}) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Only {cf_count}/3 CF thresholds found (<=30={has_30}, <=60={has_60}, <=90={has_90})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Sort order — rows sorted ascending by Expiry Date (D column) / Days Remaining (0.10 points)
    # The task says "Sort by column E ascending (soonest expiry first)"
    # Since E contains formulas and D is the underlying expiry date, we verify D is ascending.
    try:
        expiry_dates = []
        for row in range(2, 82):
            val = ws.cell(row=row, column=4).value
            expiry_dates.append(val)

        # Check if all dates are non-None and ascending
        valid_dates = [d for d in expiry_dates if d is not None]
        if len(valid_dates) == 80:
            is_sorted = all(valid_dates[i] <= valid_dates[i + 1] for i in range(len(valid_dates) - 1))
            if is_sorted:
                print(f"PASS: Component 6 — 80 rows sorted ascending by Expiry Date (soonest first) (0.10 pts)")
                total_score += 0.10
            else:
                # Find first out-of-order pair
                for i in range(len(valid_dates) - 1):
                    if valid_dates[i] > valid_dates[i + 1]:
                        print(f"FAIL: Component 6 — Not sorted ascending. Row {i+2} date {valid_dates[i]} > row {i+3} date {valid_dates[i+1]}")
                        break
        else:
            print(f"FAIL: Component 6 — Only {len(valid_dates)}/80 expiry dates are non-None")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
