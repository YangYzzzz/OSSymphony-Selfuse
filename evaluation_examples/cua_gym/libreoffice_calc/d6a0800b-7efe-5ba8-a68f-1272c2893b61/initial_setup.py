"""
Initial Setup: Contract management sheet with date-based metrics
Task ID: calc_gen_dateformulas_062
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date, timedelta
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_dateformulas_062'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contracts'

    # --- Headers ---
    headers = ['Contract ID', 'Client', 'Signed Date', 'Expiry Date',
               'Days Remaining', 'Renewal Window', 'Expiry Quarter', 'Business Days']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Realistic client names
    clients = [
        'Apex Technologies', 'BlueSky Analytics', 'Cascade Innovations', 'Delta Systems',
        'Envision Corp', 'Frontier Labs', 'GlobalEdge Solutions', 'Horizon Dynamics',
        'Intellex Partners', 'JetStream Ventures', 'Keystone Digital', 'Luminary Group',
        'Momentum Media', 'Nexus Consulting', 'Orbit Software', 'Pinnacle Strategies',
        'Quantum Networks', 'Radiant Technologies', 'Stellar Enterprises', 'Titan Analytics',
        'UniCore Systems', 'Vantage Point Inc', 'Wavefront Solutions', 'Xenith Partners',
        'Yellowstone Group', 'Zenith Innovations', 'Argon Capital', 'Beacon Analytics',
        'Cobalt Systems', 'Driftwood Media'
    ]

    # Reference today for realistic date ranges
    today = date(2026, 3, 4)  # Fixed reference date for reproducibility

    contracts = []
    for i in range(80):
        contract_id = f'CTR-2024-{str(i+1).zfill(3)}'
        client = clients[i % len(clients)]

        # Signed dates: ranging from 2023-01-01 to 2025-06-30
        signed_offset = random.randint(0, 730)
        signed_date = date(2023, 1, 1) + timedelta(days=signed_offset)

        # Expiry dates: 1 to 3 years after signing, spread across a range
        # Include some contracts expiring soon (within 30, 60, 90 days of today)
        if i < 10:
            # Expiring very soon (within 0-30 days)
            expiry_date = today + timedelta(days=random.randint(1, 30))
        elif i < 20:
            # Expiring in renewal window (31-60 days)
            expiry_date = today + timedelta(days=random.randint(31, 60))
        elif i < 30:
            # Expiring soon (61-90 days)
            expiry_date = today + timedelta(days=random.randint(61, 90))
        else:
            # Future expiry (91 days to 2 years out)
            expiry_date = today + timedelta(days=random.randint(91, 730))

        # Make sure signed_date < expiry_date
        if signed_date >= expiry_date:
            signed_date = expiry_date - timedelta(days=random.randint(180, 500))
            if signed_date < date(2022, 1, 1):
                signed_date = date(2022, 1, 1)

        contracts.append((contract_id, client, signed_date, expiry_date))

    # Shuffle to make sorting meaningful later
    random.shuffle(contracts)

    for r, (contract_id, client, signed_date, expiry_date) in enumerate(contracts, 2):
        ws.cell(row=r, column=1, value=contract_id)
        ws.cell(row=r, column=2, value=client)
        # Write dates as actual date values
        ws.cell(row=r, column=3, value=signed_date)
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=4, value=expiry_date)
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'
        # E, F, G, H intentionally left empty - will be filled by agent

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Contracts')
    print(f'  Rows: 80 contracts (rows 2-81)')
    print(f'  Columns A-D populated, E-H empty')

create_initial()
