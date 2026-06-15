"""
Initial Setup: Apply custom formula validation to cell E2
Task ID: calc_nrv_054
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Projects"

    # --- Headers ---
    headers = ['Project Name', 'Manager', 'Budget ($)', 'Start Date', 'End Date', 'Status']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    # --- Data rows ---
    # Column D (4) = Start Date, Column E (5) = End Date
    # E2 must be EMPTY with NO validation (task requires adding validation)
    data = [
        ['Website Redesign',      'Sarah Chen',       45000,  date(2024, 3, 15),  None,               'Planning'],
        ['Mobile App v2',         'Marcus Johnson',   120000, date(2024, 1, 10),  date(2024, 6, 30),  'In Progress'],
        ['Data Migration',        'Priya Patel',      78000,  date(2024, 2, 1),   date(2024, 5, 15),  'In Progress'],
        ['Cloud Infrastructure',  'James Wilson',     250000, date(2024, 4, 1),   date(2024, 12, 31), 'Planning'],
        ['Security Audit',        'Elena Rodriguez',  35000,  date(2024, 3, 20),  date(2024, 4, 30),  'Completed'],
        ['CRM Integration',       'David Kim',        92000,  date(2024, 5, 1),   date(2024, 9, 15),  'Not Started'],
        ['Analytics Dashboard',   'Lisa Wang',        67000,  date(2024, 2, 15),  date(2024, 7, 31),  'In Progress'],
        ['API Gateway',           'Robert Taylor',    88000,  date(2024, 6, 1),   date(2024, 11, 30), 'Not Started'],
        ['Employee Portal',       'Amanda Foster',    54000,  date(2024, 3, 1),   date(2024, 8, 15),  'Planning'],
        ['Inventory System',      'Carlos Mendez',    110000, date(2024, 1, 15),  date(2024, 10, 31), 'In Progress'],
        ['Payment Processing',    'Nina Kowalski',    145000, date(2024, 4, 15),  date(2024, 12, 15), 'Planning'],
        ['DevOps Pipeline',       'Tom Henderson',    63000,  date(2024, 2, 20),  date(2024, 6, 15),  'Completed'],
    ]

    date_fmt = 'yyyy-mm-dd'

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c in (4, 5) and val is not None:
                cell.number_format = date_fmt

    # Set column widths for readability
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Number format for budget column
    for r in range(2, 14):
        ws.cell(row=r, column=3).number_format = '$#,##0'

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Total Projects"
    ws2["B1"] = 12
    ws2["A2"] = "Total Budget"
    ws2["B2"] = "=SUM(Projects!C2:C13)"
    ws2["B2"].number_format = '$#,##0'
    ws2["A1"].font = Font(bold=True)
    ws2["A2"].font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
