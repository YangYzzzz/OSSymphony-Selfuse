"""
Initial Setup: Create spreadsheet with data for macro filtering task
Task ID: calc_mcp_019
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers in A1:F1
    headers = ["Employee ID", "Name", "Department", "Category", "Amount", "Date"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic employee data - 99 rows (A2:F100)
    first_names = [
        "Sarah", "Marcus", "Emily", "James", "Priya", "Carlos", "Aisha",
        "David", "Mei", "Robert", "Fatima", "Thomas", "Yuki", "Daniel",
        "Olivia", "Nathan", "Sofia", "Kevin", "Rachel", "Andre",
        "Hannah", "Luis", "Emma", "Michael", "Zara", "Patrick",
        "Isabella", "Brian", "Chloe", "Vincent"
    ]
    last_names = [
        "Chen", "Johnson", "Williams", "Brown", "Patel", "Garcia", "Ahmed",
        "Taylor", "Zhang", "Anderson", "Hassan", "Moore", "Tanaka", "Kim",
        "Martinez", "Wilson", "Lopez", "Lee", "Walker", "Dubois",
        "Scott", "Rivera", "Torres", "White", "Khan", "Murphy",
        "Rossi", "Clark", "Bennett", "Nguyen"
    ]
    departments = [
        "Engineering", "Marketing", "Sales", "Finance", "Human Resources",
        "Operations", "Research", "Customer Support", "Legal", "Product"
    ]
    categories = [
        "Travel", "Equipment", "Software", "Training", "Office Supplies",
        "Consulting", "Maintenance", "Catering", "Subscriptions", "Shipping"
    ]

    # Generate amounts with a good mix above and below 500
    # Roughly 40-50 rows above 500, rest below
    amounts = []
    for i in range(99):
        if random.random() < 0.45:
            # Below or equal to 500
            amounts.append(round(random.uniform(50, 500), 2))
        else:
            # Above 500
            amounts.append(round(random.uniform(501, 5000), 2))

    months = list(range(1, 13))
    days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]

    for i in range(99):
        row = i + 2
        emp_id = f"EMP-{1001 + i:04d}"
        first = random.choice(first_names)
        last = random.choice(last_names)
        name = f"{first} {last}"
        dept = random.choice(departments)
        cat = random.choice(categories)
        amount = amounts[i]
        month = random.choice(months)
        day = random.randint(1, days_in_month[month - 1])
        date_str = f"2025-{month:02d}-{day:02d}"

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=dept)
        ws.cell(row=row, column=4, value=cat)
        ws.cell(row=row, column=5, value=amount)
        ws.cell(row=row, column=6, value=date_str)

    # Set reasonable column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
