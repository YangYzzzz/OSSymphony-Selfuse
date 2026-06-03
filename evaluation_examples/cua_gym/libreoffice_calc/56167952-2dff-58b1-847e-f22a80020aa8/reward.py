"""
Reward Script: Verify 'GenerateReport' macro output
Task ID: calc_mcp_019
Domain: libreoffice_calc
Scoring:
  Component 1: 'Report' sheet exists (0.15)
  Component 2: Report headers match Sheet1 headers A1:F1 (0.20)
  Component 3: All Report data rows have column E > 500 (0.25)
  Component 4: Correct row count matches Sheet1 filtered count (0.20)
  Component 5: Row order preserved from Sheet1 (0.20)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_019'


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Sheet1 must exist with data
    if 'Sheet1' not in wb.sheetnames:
        print("CRITICAL: Sheet1 not found — file is corrupted or wrong")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']

    # Component 1: 'Report' sheet exists (0.15 points)
    try:
        if 'Report' in wb.sheetnames:
            print(f"PASS: Component 1 — 'Report' sheet exists (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — 'Report' sheet does not exist")
            # Without Report sheet, no further checks can pass
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    rws = wb['Report']

    # Collect Sheet1 headers
    s1_headers = [ws1.cell(row=1, column=c).value for c in range(1, 7)]

    # Component 2: Report headers match Sheet1 headers A1:F1 (0.20 points)
    try:
        report_headers = [rws.cell(row=1, column=c).value for c in range(1, 7)]
        if report_headers == s1_headers:
            print(f"PASS: Component 2 — Report headers match Sheet1 headers: {report_headers} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Headers mismatch. Expected {s1_headers}, found {report_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Collect Sheet1 rows where column E > 500 (preserving order)
    s1_filtered_rows = []
    for r in range(2, ws1.max_row + 1):
        e_val = ws1.cell(row=r, column=5).value
        if e_val is not None and isinstance(e_val, (int, float)) and e_val > 500:
            row_data = tuple(ws1.cell(row=r, column=c).value for c in range(1, 7))
            s1_filtered_rows.append(row_data)

    # Collect Report data rows (skip header row 1)
    report_data_rows = []
    for r in range(2, rws.max_row + 1):
        row_data = tuple(rws.cell(row=r, column=c).value for c in range(1, 7))
        # Skip completely empty rows
        if any(v is not None for v in row_data):
            report_data_rows.append(row_data)

    # Component 3: All Report data rows have column E > 500 (0.25 points)
    try:
        if len(report_data_rows) == 0:
            print(f"FAIL: Component 3 — Report has no data rows")
        else:
            bad_rows = sum(
                1 for row in report_data_rows
                if row[4] is None or not isinstance(row[4], (int, float)) or row[4] <= 500
            )
            if bad_rows == 0:
                print(f"PASS: Component 3 — All {len(report_data_rows)} Report rows have E > 500 (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — {bad_rows}/{len(report_data_rows)} rows have E <= 500 or invalid")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Correct row count (0.20 points)
    try:
        expected_count = len(s1_filtered_rows)
        actual_count = len(report_data_rows)
        if actual_count == expected_count:
            print(f"PASS: Component 4 — Report has {actual_count} rows matching Sheet1 filtered count (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Expected {expected_count} rows, found {actual_count}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Row order preserved from Sheet1 (0.20 points)
    try:
        if len(report_data_rows) == 0:
            print(f"FAIL: Component 5 — No report rows to check order")
        elif report_data_rows == s1_filtered_rows:
            print(f"PASS: Component 5 — Row order matches Sheet1 filtered order (0.20 pts)")
            total_score += 0.20
        else:
            # Check how many rows match in order
            matches = 0
            check_len = min(len(report_data_rows), len(s1_filtered_rows))
            for i in range(check_len):
                if report_data_rows[i] == s1_filtered_rows[i]:
                    matches += 1
            print(f"FAIL: Component 5 — Order mismatch. {matches}/{check_len} rows in correct position")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
