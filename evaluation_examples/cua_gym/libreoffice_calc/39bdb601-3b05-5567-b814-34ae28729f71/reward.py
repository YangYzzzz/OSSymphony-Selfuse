"""
Reward Script: Fix VLOOKUP to retrieve second occurrence of 'OrderID-500'
Task ID: calc_tbl_042
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Formulas in B6:B10 are no longer simple VLOOKUP
  Component 2 (0.4): Formulas structurally target the 2nd occurrence
  Component 3 (0.3): All 5 lookup cells contain consistent replacement formulas
"""

import os
import re
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_042'

# Known data for the second occurrence of OrderID-500 (row 42 in Orders sheet):
# Customer: Daniel Clark, Product: Ergonomic Chair Pro, Amount: 349.5,
# Date: 2025-02-12, Status: Delivered

def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Precondition: B3 must still reference OrderID-500
    lookup_id = ws['B3'].value
    if lookup_id != 'OrderID-500':
        print(f"CRITICAL: B3 lookup ID changed from 'OrderID-500' to '{lookup_id}'")
        print("REWARD: 0.0")
        return 0.0

    # Collect formulas from B6:B10
    formula_cells = {}
    for row in range(6, 11):
        cell_val = ws.cell(row=row, column=2).value
        formula_cells[row] = cell_val
        label = ws.cell(row=row, column=1).value
        print(f"  B{row} ({label}): {cell_val}")

    # Component 1: Formulas in B6:B10 are NO LONGER simple VLOOKUP (0.3 points)
    # Initial state uses =VLOOKUP(B3,Orders!A:F,N,FALSE)
    # If any cell still uses plain VLOOKUP, it returns the 1st match (wrong)
    try:
        vlookup_count = 0
        formula_count = 0
        for row in range(6, 11):
            val = formula_cells[row]
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_count += 1
                # Check if it's a simple VLOOKUP (not containing INDEX, AGGREGATE, MATCH, SMALL, COUNTIF etc.)
                upper_val = val.upper().replace(" ", "")
                if upper_val.startswith('=VLOOKUP(') and 'INDEX' not in upper_val and 'AGGREGATE' not in upper_val and 'MATCH' not in upper_val and 'SMALL' not in upper_val and 'COUNTIF' not in upper_val:
                    vlookup_count += 1

        if formula_count >= 5 and vlookup_count == 0:
            print(f"PASS: Component 1 - All 5 formulas replaced from simple VLOOKUP (0.3 pts)")
            total_score += 0.3
        elif formula_count >= 5 and vlookup_count < 5:
            partial = 0.3 * (5 - vlookup_count) / 5
            print(f"PARTIAL: Component 1 - {5 - vlookup_count}/5 formulas replaced ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Found {vlookup_count} simple VLOOKUPs out of {formula_count} formulas")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Formulas structurally target the 2nd occurrence (0.4 points)
    # Valid approaches include:
    # - INDEX + AGGREGATE(15,6,...,2) where the trailing ,2 means 2nd smallest
    # - INDEX + MATCH + COUNTIF combination
    # - INDEX + SMALL + IF array formula
    # - OFFSET-based approach
    # - Any formula that explicitly references "2" as the occurrence number
    try:
        second_occ_count = 0
        for row in range(6, 11):
            val = formula_cells[row]
            if val is None or not isinstance(val, str) or not val.startswith('='):
                continue
            upper_val = val.upper().replace(" ", "")

            # Check multiple patterns that indicate 2nd-occurrence retrieval
            # Pattern A: AGGREGATE(15,6,...,2) - the ,2) at end means 2nd smallest
            pat_a = 'AGGREGATE' in upper_val and re.search(r',2\)', upper_val)
            # Pattern B: SMALL(...,2) - 2nd smallest
            pat_b = 'SMALL(' in upper_val and re.search(r'SMALL\([^)]*,\s*2\)', upper_val)
            # Pattern C: COUNTIF-based offset for 2nd match
            pat_c = 'COUNTIF' in upper_val and ('MATCH' in upper_val or 'INDEX' in upper_val)
            # Pattern D: Explicit row reference to row 42 (the 2nd occurrence row)
            pat_d = bool(re.search(r'ORDERS![A-F]\$?42', upper_val, re.IGNORECASE))
            # Pattern E: INDEX with ",2)" (common nth-occurrence pattern)
            pat_e = 'INDEX' in upper_val and ',2)' in upper_val

            if pat_a or pat_b or pat_c or pat_d or pat_e:
                second_occ_count += 1

        if second_occ_count >= 5:
            print(f"PASS: Component 2 - All 5 formulas target 2nd occurrence (0.4 pts)")
            total_score += 0.4
        elif second_occ_count > 0:
            partial = 0.4 * second_occ_count / 5
            print(f"PARTIAL: Component 2 - {second_occ_count}/5 formulas target 2nd occurrence ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No formulas appear to target the 2nd occurrence")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: All 5 lookup cells have consistent, well-formed replacement formulas (0.3 points)
    # Each formula should reference the Orders sheet and the correct column for its field:
    #   B6 (Customer) -> column B (col 2)
    #   B7 (Product)  -> column C (col 3)
    #   B8 (Amount)   -> column D (col 4)
    #   B9 (Date)     -> column E (col 5)
    #   B10 (Status)  -> column F (col 6)
    try:
        expected_cols = {
            6: 'B',   # Customer
            7: 'C',   # Product
            8: 'D',   # Amount
            9: 'E',   # Date
            10: 'F',  # Status
        }
        consistent_count = 0
        for row, expected_col in expected_cols.items():
            val = formula_cells[row]
            if val is None or not isinstance(val, str) or not val.startswith('='):
                print(f"  B{row}: not a formula")
                continue
            upper_val = val.upper()
            # Check that formula references the correct Orders column
            # Patterns: Orders!B, Orders!$B, ORDERS!B$2:B$51, etc.
            col_pattern = f"ORDERS!{expected_col}" if expected_col != 'B' else "ORDERS!B"
            # Also handle $-prefixed references
            col_pattern_dollar = f"ORDERS!\\${expected_col}"
            if re.search(f"ORDERS!\\$?{expected_col}", upper_val):
                consistent_count += 1
            else:
                print(f"  B{row}: expected reference to Orders!{expected_col}, formula: {val}")

        if consistent_count >= 5:
            print(f"PASS: Component 3 - All 5 formulas reference correct Orders columns (0.3 pts)")
            total_score += 0.3
        elif consistent_count > 0:
            partial = 0.3 * consistent_count / 5
            print(f"PARTIAL: Component 3 - {consistent_count}/5 formulas reference correct columns ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No formulas reference correct Orders columns")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
