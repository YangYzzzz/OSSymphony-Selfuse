"""
Initial Setup: Lookup formula returns wrong value due to duplicate keys
Task ID: calc_tbl_042
Domain: libreoffice_calc

Creates an Orders lookup table with duplicate OrderID-500 in rows 15 and 42,
and a Summary sheet with a VLOOKUP that incorrectly returns the first match.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Orders (lookup table) ---
    ws = wb.active
    ws.title = "Orders"

    headers = ["OrderID", "Customer", "Product", "Amount", "Date", "Status"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Generate 50 rows of realistic order data
    # OrderID-500 appears in row 15 (original) and row 42 (revised)
    order_data = [
        # row 2-14: orders before the first OrderID-500
        ("OrderID-488", "Liam Parker", "Wireless Mouse", 29.99, "2025-01-03", "Shipped"),
        ("OrderID-489", "Sophia Martinez", "USB-C Hub", 45.50, "2025-01-04", "Delivered"),
        ("OrderID-490", "Noah Williams", "Mechanical Keyboard", 89.00, "2025-01-05", "Shipped"),
        ("OrderID-491", "Emma Brown", "Monitor Stand", 34.75, "2025-01-06", "Processing"),
        ("OrderID-492", "James Davis", "Webcam HD", 62.00, "2025-01-07", "Delivered"),
        ("OrderID-493", "Olivia Wilson", "Laptop Sleeve 15in", 24.99, "2025-01-08", "Shipped"),
        ("OrderID-494", "Benjamin Taylor", "HDMI Cable 6ft", 12.49, "2025-01-09", "Delivered"),
        ("OrderID-495", "Ava Anderson", "Desk Lamp LED", 38.00, "2025-01-10", "Shipped"),
        ("OrderID-496", "Lucas Thomas", "Mousepad XL", 19.95, "2025-01-11", "Delivered"),
        ("OrderID-497", "Mia Jackson", "Bluetooth Speaker", 55.00, "2025-01-12", "Processing"),
        ("OrderID-498", "Ethan White", "Phone Stand", 15.99, "2025-01-13", "Shipped"),
        ("OrderID-499", "Isabella Harris", "Power Strip 6-Outlet", 22.50, "2025-01-14", "Delivered"),
        # row 15: FIRST occurrence of OrderID-500 (original order)
        ("OrderID-500", "Daniel Clark", "Ergonomic Chair", 299.00, "2025-01-15", "Cancelled"),
        # row 16-41: more orders
        ("OrderID-501", "Charlotte Lewis", "Standing Desk Mat", 42.00, "2025-01-16", "Shipped"),
        ("OrderID-502", "Alexander Robinson", "Cable Management Kit", 18.75, "2025-01-17", "Delivered"),
        ("OrderID-503", "Amelia Walker", "Desk Organizer", 27.50, "2025-01-18", "Shipped"),
        ("OrderID-504", "Henry Hall", "Noise-Cancel Headphones", 179.99, "2025-01-19", "Delivered"),
        ("OrderID-505", "Harper Allen", "Portable Charger 20K", 35.00, "2025-01-20", "Processing"),
        ("OrderID-506", "Sebastian Young", "USB Flash Drive 128GB", 14.99, "2025-01-21", "Shipped"),
        ("OrderID-507", "Evelyn King", "Laptop Cooling Pad", 28.50, "2025-01-22", "Delivered"),
        ("OrderID-508", "Jack Wright", "Wireless Earbuds", 49.99, "2025-01-23", "Shipped"),
        ("OrderID-509", "Abigail Scott", "Screen Protector Pack", 9.99, "2025-01-24", "Delivered"),
        ("OrderID-510", "Owen Green", "External SSD 1TB", 89.99, "2025-01-25", "Processing"),
        ("OrderID-511", "Emily Adams", "Keyboard Wrist Rest", 16.50, "2025-01-26", "Shipped"),
        ("OrderID-512", "Michael Baker", "HDMI Adapter", 11.99, "2025-01-27", "Delivered"),
        ("OrderID-513", "Ella Gonzalez", "Ring Light 10in", 32.00, "2025-01-28", "Shipped"),
        ("OrderID-514", "William Nelson", "Surge Protector", 26.75, "2025-01-29", "Delivered"),
        ("OrderID-515", "Scarlett Carter", "Webcam Cover 3-Pack", 5.99, "2025-01-30", "Shipped"),
        ("OrderID-516", "Daniel Mitchell", "Smart Plug 4-Pack", 39.99, "2025-01-31", "Processing"),
        ("OrderID-517", "Grace Perez", "Document Scanner", 149.00, "2025-02-01", "Delivered"),
        ("OrderID-518", "Matthew Roberts", "USB Microphone", 67.50, "2025-02-02", "Shipped"),
        ("OrderID-519", "Chloe Turner", "Wireless Trackpad", 59.00, "2025-02-03", "Delivered"),
        ("OrderID-520", "David Phillips", "Monitor Arm", 44.99, "2025-02-04", "Shipped"),
        ("OrderID-521", "Lily Campbell", "Ethernet Cable 25ft", 8.99, "2025-02-05", "Delivered"),
        ("OrderID-522", "Joseph Parker", "Docking Station", 129.00, "2025-02-06", "Processing"),
        ("OrderID-523", "Zoey Edwards", "Portable Monitor 15in", 199.00, "2025-02-07", "Shipped"),
        ("OrderID-524", "Andrew Collins", "Keyboard Cleaner", 7.99, "2025-02-08", "Delivered"),
        ("OrderID-525", "Penelope Stewart", "Laptop Backpack", 54.99, "2025-02-09", "Shipped"),
        ("OrderID-526", "Christopher Sanchez", "WiFi Range Extender", 33.00, "2025-02-10", "Delivered"),
        ("OrderID-527", "Nora Morris", "Presentation Remote", 22.99, "2025-02-11", "Shipped"),
        # row 42: SECOND occurrence of OrderID-500 (revised order)
        ("OrderID-500", "Daniel Clark", "Ergonomic Chair Pro", 349.50, "2025-02-12", "Delivered"),
        # row 43-51: remaining orders
        ("OrderID-528", "Ryan Rogers", "USB-C Charger 65W", 29.99, "2025-02-13", "Shipped"),
        ("OrderID-529", "Hannah Reed", "Desk Shelf Riser", 37.50, "2025-02-14", "Delivered"),
        ("OrderID-530", "Nathan Cook", "Cable Clips 20-Pack", 6.49, "2025-02-15", "Shipped"),
        ("OrderID-531", "Aria Morgan", "Bluetooth Mouse", 24.99, "2025-02-16", "Processing"),
        ("OrderID-532", "Samuel Bell", "Mini Projector", 159.00, "2025-02-17", "Shipped"),
        ("OrderID-533", "Victoria Murphy", "Desk Calendar 2025", 12.99, "2025-02-18", "Delivered"),
        ("OrderID-534", "Dylan Bailey", "USB Hub 7-Port", 21.50, "2025-02-19", "Shipped"),
        ("OrderID-535", "Aurora Rivera", "Laptop Stand Aluminum", 46.00, "2025-02-20", "Delivered"),
        ("OrderID-536", "Luke Cooper", "Wireless Charger Pad", 19.99, "2025-02-21", "Shipped"),
    ]

    for r_idx, row_data in enumerate(order_data, 2):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            if c_idx == 4:  # Amount column
                cell.number_format = '$#,##0.00'
            elif c_idx == 5:  # Date column
                cell.number_format = 'yyyy-mm-dd'

    # Set column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet("Summary")

    # Title
    ws2["A1"] = "Order Lookup Summary"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True)

    ws2["A3"] = "Lookup OrderID:"
    ws2["A3"].font = Font(bold=True)
    ws2["B3"] = "OrderID-500"

    ws2["A5"] = "Field"
    ws2["B5"] = "Value (from VLOOKUP)"
    ws2["A5"].font = Font(bold=True)
    ws2["B5"].font = Font(bold=True)

    # VLOOKUP formulas that return FIRST occurrence (row 15 data) - this is the BUG
    lookup_fields = [
        ("Customer", 2),
        ("Product", 3),
        ("Amount", 4),
        ("Date", 5),
        ("Status", 6),
    ]

    for i, (field_name, col_index) in enumerate(lookup_fields):
        row = 6 + i
        ws2.cell(row=row, column=1, value=field_name)
        ws2.cell(row=row, column=1).font = Font(bold=True)
        # VLOOKUP returns the first match (row 15 - Cancelled order)
        ws2.cell(row=row, column=2,
                 value=f'=VLOOKUP(B3,Orders!A:F,{col_index},FALSE)')

    # Note about the issue
    ws2["A12"] = "Note:"
    ws2["A12"].font = Font(bold=True, color="FF0000")
    ws2["B12"] = "The VLOOKUP above returns data from the original order (row 15), but we need the revised order (row 42)."
    ws2["B12"].font = Font(color="FF0000")

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
