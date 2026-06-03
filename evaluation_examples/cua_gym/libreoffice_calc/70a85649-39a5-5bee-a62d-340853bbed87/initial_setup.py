"""
Initial Setup: Create a spreadsheet with headers in A1:L1 and 499 rows of employee data.
Task ID: calc_nrv_042
Domain: libreoffice_calc
No named ranges, no print title rows.
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_042'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employees"

    # Headers in A1:L1
    headers = [
        'ID', 'Name', 'Department', 'Hire Date', 'Salary',
        'Bonus', 'Total', 'Status', 'Location', 'Manager',
        'Rating', 'Notes'
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic data pools
    first_names = [
        'Sarah', 'Marcus', 'Priya', 'James', 'Olivia', 'Wei', 'Carlos',
        'Aisha', 'Thomas', 'Yuki', 'Elena', 'David', 'Fatima', 'Robert',
        'Mei', 'Alexander', 'Zara', 'Benjamin', 'Sophia', 'Liam',
        'Amara', 'Noah', 'Isabella', 'Ethan', 'Ava', 'Lucas', 'Mia',
        'Daniel', 'Charlotte', 'Henry', 'Chloe', 'Raj', 'Emily', 'Omar'
    ]
    last_names = [
        'Chen', 'Johnson', 'Patel', 'Williams', 'Kim', 'Garcia', 'Brown',
        'Singh', 'Miller', 'Tanaka', 'Rodriguez', 'Anderson', 'Taylor',
        'Lee', 'Martinez', 'Wilson', 'Moore', 'Clark', 'Hall', 'Davis',
        'Lopez', 'Young', 'King', 'Wright', 'Scott', 'Baker', 'Adams',
        'Nelson', 'Hill', 'Campbell', 'Mitchell', 'Roberts', 'Carter'
    ]
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Finance', 'Human Resources',
        'Operations', 'Legal', 'Product', 'Customer Support', 'Research'
    ]
    statuses = ['Active', 'On Leave', 'Probation', 'Active', 'Active', 'Active']
    locations = [
        'New York', 'San Francisco', 'Chicago', 'Austin', 'Seattle',
        'Boston', 'Denver', 'Atlanta', 'Portland', 'Miami'
    ]
    managers = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James Williams',
        'Olivia Kim', 'Wei Garcia', 'Carlos Brown', 'Aisha Singh',
        'Thomas Miller', 'Yuki Tanaka'
    ]
    notes_pool = [
        'Top performer Q4', 'Transferred from London office',
        'Leading cloud migration project', 'Mentoring junior team members',
        'Completed leadership training', 'Working on AI initiative',
        'Cross-functional team lead', 'Remote work arrangement',
        'Patent pending for new algorithm', 'Bilingual - Spanish/English',
        'MBA from Stanford', 'Former startup founder',
        'Specializes in data analytics', 'Key account manager',
        'Certified Scrum Master', '', '', '', ''
    ]

    # Generate 499 rows of data (rows 2-500)
    for r in range(2, 501):
        emp_id = f'EMP-{r - 1:04d}'
        name = f'{random.choice(first_names)} {random.choice(last_names)}'
        dept = random.choice(departments)
        # Hire dates between 2018-01-01 and 2025-12-31
        year = random.randint(2018, 2025)
        month = random.randint(1, 12)
        day = random.randint(1, 28)
        hire_date = f'{year}-{month:02d}-{day:02d}'
        salary = round(random.randint(45000, 185000), -2)
        bonus = round(salary * random.uniform(0.03, 0.20), 2)
        total = salary + bonus
        status = random.choice(statuses)
        location = random.choice(locations)
        manager = random.choice(managers)
        rating = round(random.uniform(2.5, 5.0), 1)
        notes = random.choice(notes_pool)

        row_data = [
            emp_id, name, dept, hire_date, salary,
            bonus, total, status, location, manager,
            rating, notes
        ]
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    col_widths = {'A': 12, 'B': 22, 'C': 18, 'D': 14, 'E': 12,
                  'F': 12, 'G': 12, 'H': 12, 'I': 16, 'J': 20,
                  'K': 10, 'L': 35}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
