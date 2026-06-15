"""
Initial Setup: Survey response spreadsheet with empty rating columns
Task ID: calc_gcv_082
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Survey_Response'

    # Headers
    headers = ['Respondent ID', 'Question', 'Rating Text', 'Rating Number']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 44 survey questions across various categories
    questions = [
        'How satisfied are you with the overall onboarding experience?',
        'How would you rate the clarity of job role expectations?',
        'How effective was the initial training program?',
        'How accessible is your direct supervisor for questions?',
        'How satisfied are you with the team collaboration tools?',
        'How would you rate the quality of internal documentation?',
        'How comfortable do you feel raising concerns to management?',
        'How satisfied are you with the office workspace environment?',
        'How would you rate the IT support responsiveness?',
        'How effective are the weekly team meetings?',
        'How satisfied are you with the performance review process?',
        'How would you rate opportunities for professional development?',
        'How clear are the company goals and strategic direction?',
        'How satisfied are you with the employee benefits package?',
        'How would you rate the work-life balance at this company?',
        'How effective is cross-department communication?',
        'How satisfied are you with the recognition and rewards program?',
        'How would you rate the fairness of the promotion process?',
        'How comfortable are you with the company culture?',
        'How satisfied are you with the diversity and inclusion efforts?',
        'How would you rate the quality of leadership communication?',
        'How effective is the project management methodology?',
        'How satisfied are you with the remote work policy?',
        'How would you rate the company social events?',
        'How clear are the expectations for your current project?',
        'How satisfied are you with the mentorship opportunities?',
        'How would you rate the conflict resolution process?',
        'How effective is the feedback mechanism in your team?',
        'How satisfied are you with the salary review frequency?',
        'How would you rate the health and wellness programs?',
        'How comfortable are you using the company intranet?',
        'How satisfied are you with the parking and commute support?',
        'How would you rate the cafeteria and food options?',
        'How effective is the customer feedback integration process?',
        'How satisfied are you with the data security training?',
        'How would you rate the software tools provided for your role?',
        'How comfortable are you with the change management process?',
        'How satisfied are you with the sprint planning sessions?',
        'How would you rate the quality of the code review process?',
        'How effective is the incident response procedure?',
        'How satisfied are you with the vacation and leave policy?',
        'How would you rate the accessibility of HR services?',
        'How comfortable are you providing anonymous feedback?',
        'How satisfied are you with the overall company direction?',
    ]

    for i, question in enumerate(questions):
        row = i + 2
        ws.cell(row=row, column=1, value=f'RESP-{1001 + i}')
        ws.cell(row=row, column=2, value=question)
        # C and D columns left empty — no validation applied

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 55
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
