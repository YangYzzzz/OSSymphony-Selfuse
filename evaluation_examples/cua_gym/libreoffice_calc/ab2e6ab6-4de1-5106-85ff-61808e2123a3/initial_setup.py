"""
Initial Setup: Apply 'Neutral' built-in cell style to projected values
Task ID: calc_gfl_077
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_077'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Projections ---
    ws = wb.active
    ws.title = 'Projections'

    # Headers (row 1)
    headers = ['Category', 'Q1 Forecast ($K)', 'Q2 Forecast ($K)', 'Q3 Forecast ($K)']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font_white = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    # Rows 2-4: Historical data (actual values)
    historical_data = [
        ['Revenue - North America', 1245.8, 1312.4, 1389.1],
        ['Revenue - Europe', 876.3, 921.5, 958.7],
        ['Revenue - Asia Pacific', 534.2, 567.8, 612.3],
    ]
    for r, row_data in enumerate(historical_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '#,##0.0'

    # Rows 5-15: Projected values (estimates) - NO special styling
    projected_data = [
        ['Operating Expenses', 487.3, 502.1, 518.6],
        ['Marketing Budget', 156.9, 163.4, 171.2],
        ['R&D Investment', 234.5, 248.7, 261.3],
        ['Personnel Costs', 612.8, 631.4, 649.7],
        ['Infrastructure', 89.4, 92.1, 95.8],
        ['Client Acquisition', 178.3, 185.9, 194.2],
        ['Product Development', 321.6, 338.2, 355.7],
        ['Supply Chain', 145.7, 151.3, 157.9],
        ['Legal & Compliance', 67.2, 69.8, 72.5],
        ['Training & Development', 43.1, 45.6, 48.2],
        ['Contingency Reserve', 112.4, 118.9, 125.3],
    ]
    for r, row_data in enumerate(projected_data, 5):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c >= 2:
                cell.number_format = '#,##0.0'

    # Rows 16-20: Summary rows
    summary_data = [
        ['Total Projected Expenses', None, None, None],
        ['Average per Quarter', None, None, None],
        ['Year-over-Year Growth %', None, None, None],
        ['Budget Variance', None, None, None],
        ['Net Forecast Position', None, None, None],
    ]
    # Fill summary with formulas
    summary_values = [
        ['Total Projected Expenses', '=SUM(B5:B15)', '=SUM(C5:C15)', '=SUM(D5:D15)'],
        ['Average per Quarter', '=AVERAGE(B5:B15)', '=AVERAGE(C5:C15)', '=AVERAGE(D5:D15)'],
        ['Year-over-Year Growth %', 8.2, 9.1, 10.4],
        ['Budget Variance', -23.5, -18.7, -12.3],
        ['Net Forecast Position', 1456.2, 1523.8, 1601.4],
    ]
    summary_font = Font(name='Calibri', size=11, bold=True)
    for r, row_data in enumerate(summary_values, 16):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.font = summary_font
            elif c >= 2 and isinstance(val, (int, float)):
                cell.number_format = '#,##0.0'

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
