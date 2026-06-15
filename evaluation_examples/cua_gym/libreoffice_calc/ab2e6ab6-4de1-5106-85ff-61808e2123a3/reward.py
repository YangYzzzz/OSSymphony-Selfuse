"""
Reward Script: Apply 'Neutral' built-in cell style to cells B5:D15
Task ID: calc_gfl_077
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): B5:D15 have solid fill with Neutral style background color (FFFFEB9C)
  Component 2 (0.3): B5:D15 have Neutral style font color (009C6500)
  Component 3 (0.2): Historical rows 2-4 and summary rows 16-20 are NOT styled with Neutral fill
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_077'

# Neutral style expected colors
NEUTRAL_FILL_COLOR = 'FFFFEB9C'   # yellow/amber background (ARGB)
NEUTRAL_FONT_COLOR = '009C6500'   # dark amber font color (ARGB)


def persist_app_state(domain: str):
    """Attempt to save any unsaved LibreOffice state via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Projections' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Projections' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Projections']

    # Component 1: B5:D15 all have solid fill with Neutral background color FFFFEB9C (0.5 points)
    try:
        styled_count = 0
        total_cells = 0
        for row in range(5, 16):  # rows 5-15 inclusive
            for col in range(2, 5):  # columns B(2), C(3), D(4)
                total_cells += 1
                cell = ws.cell(row=row, column=col)
                fill_type = cell.fill.patternType
                fg_color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                if fill_type == 'solid' and fg_color == NEUTRAL_FILL_COLOR:
                    styled_count += 1

        if styled_count == total_cells:
            print(f"PASS: Component 1 - All {total_cells} cells in B5:D15 have Neutral fill color ({NEUTRAL_FILL_COLOR}) (0.5 pts)")
            total_score += 0.5
        elif styled_count > 0:
            # Partial credit: proportion of correctly styled cells
            partial = round(0.5 * (styled_count / total_cells), 2)
            print(f"PARTIAL: Component 1 - {styled_count}/{total_cells} cells have Neutral fill ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No cells in B5:D15 have Neutral fill color. Expected {NEUTRAL_FILL_COLOR}")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: B5:D15 all have Neutral font color 009C6500 (0.3 points)
    try:
        font_count = 0
        total_cells = 0
        for row in range(5, 16):
            for col in range(2, 5):
                total_cells += 1
                cell = ws.cell(row=row, column=col)
                try:
                    font_color = cell.font.color.rgb if cell.font.color else None
                except Exception:
                    font_color = None
                if font_color == NEUTRAL_FONT_COLOR:
                    font_count += 1

        if font_count == total_cells:
            print(f"PASS: Component 2 - All {total_cells} cells in B5:D15 have Neutral font color ({NEUTRAL_FONT_COLOR}) (0.3 pts)")
            total_score += 0.3
        elif font_count > 0:
            partial = round(0.3 * (font_count / total_cells), 2)
            print(f"PARTIAL: Component 2 - {font_count}/{total_cells} cells have Neutral font color ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No cells in B5:D15 have Neutral font color. Expected {NEUTRAL_FONT_COLOR}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: B5:D15 have Neutral fill AND rows 2-4, 16-20 in cols B-D are NOT styled
    #              with Neutral fill (correct scope - only target range styled) (0.2 points)
    #              This is a compound check: target cells ARE styled + non-target cells are NOT.
    #              Both parts must hold to earn the points.
    try:
        # Part A: at least one target cell has the Neutral fill (anchors to the change)
        any_target_styled = False
        for row in range(5, 16):
            for col in range(2, 5):
                cell = ws.cell(row=row, column=col)
                if cell.fill.patternType == 'solid' and cell.fill.fgColor and cell.fill.fgColor.rgb == NEUTRAL_FILL_COLOR:
                    any_target_styled = True
                    break
            if any_target_styled:
                break

        # Part B: no non-target cell in the check range has the Neutral fill
        unchanged_rows = list(range(2, 5)) + list(range(16, 21))  # rows 2-4, 16-20
        incorrectly_styled = 0
        total_check = 0
        for row in unchanged_rows:
            for col in range(2, 5):
                total_check += 1
                cell = ws.cell(row=row, column=col)
                fill_type = cell.fill.patternType
                fg_color = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                if fill_type == 'solid' and fg_color == NEUTRAL_FILL_COLOR:
                    incorrectly_styled += 1

        if any_target_styled and incorrectly_styled == 0:
            print(f"PASS: Component 3 - Neutral fill correctly scoped to B5:D15 only; {total_check} non-target cells unchanged (0.2 pts)")
            total_score += 0.2
        elif not any_target_styled:
            print(f"FAIL: Component 3 - No target cells have Neutral fill, so scope check is moot")
        else:
            print(f"FAIL: Component 3 - {incorrectly_styled}/{total_check} cells outside B5:D15 incorrectly have Neutral fill")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
