"""
Initial Setup: Fill seniority formula and create concatenated summary column
Task ID: osworld_calc_formula_pattern_concat_004
Domain: libreoffice_calc

Creates an employee spreadsheet with:
- Columns: Employee ID (A), Name (B), Department (C), Annual Salary (D), Years (E), Seniority Level (F)
- F2 has a seniority formula (IF-based on years), F3:F13 are EMPTY (agent must fill down)
- Column G header exists but G2:G13 are EMPTY (agent must add concatenation formulas)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_004'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Employees ---
    ws = wb.active
    ws.title = 'Employees'

    # Headers
    headers = ['Employee ID', 'Name', 'Department', 'Annual Salary', 'Years', 'Seniority Level', 'Summary']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Realistic employee data: (ID, Name, Department, Annual Salary, Years)
    data = [
        ('EMP001', 'Sarah Chen',        'Engineering',  85000, 7),
        ('EMP002', 'Marcus Johnson',    'Marketing',    72000, 3),
        ('EMP003', 'Priya Patel',       'Finance',      91000, 12),
        ('EMP004', 'David Williams',    'Engineering',  78500, 5),
        ('EMP005', 'Amelia Torres',     'HR',           63000, 2),
        ('EMP006', 'James Okafor',      'Operations',   69500, 8),
        ('EMP007', 'Li Wei',            'Engineering',  94000, 15),
        ('EMP008', 'Fatima Al-Hassan',  'Marketing',    67000, 4),
        ('EMP009', 'Robert Kim',        'Finance',      88000, 9),
        ('EMP010', 'Elena Vasquez',     'HR',           61500, 1),
        ('EMP011', 'Thomas Nguyen',     'Operations',   75000, 6),
        ('EMP012', 'Aisha Brown',       'Engineering',  82000, 11),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Employee ID
        ws.cell(row=r, column=2, value=row_data[1])  # Name
        ws.cell(row=r, column=3, value=row_data[2])  # Department
        ws.cell(row=r, column=4, value=row_data[3])  # Annual Salary
        ws.cell(row=r, column=5, value=row_data[4])  # Years
        # Column F: only F2 has formula; F3:F13 are intentionally left empty
        # Column G: intentionally left empty (agent must fill)

    # F2 only: seniority formula (agent must fill down to F3:F13)
    ws.cell(row=2, column=6, value='=IF(E2>=10,"Senior",IF(E2>=5,"Mid-Level","Junior"))')

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 50

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
