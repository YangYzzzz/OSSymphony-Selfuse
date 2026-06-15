"""
Reward Script: Fill Seniority formula in column F and create concatenation formula in column G
Task ID: osworld_calc_formula_pattern_concat_004
Domain: libreoffice_calc

Scoring:
  Component 1: Column F filled for all data rows (F3-F13) with correct IF-based seniority formula (0.5 pts)
  Component 2: Column G filled for all data rows (G2-G13) with correct concatenation formula (0.5 pts)
  Total: 1.0

Context:
  Initial state: F2 has an IF seniority formula; F3:F13 and G2:G13 are all empty.
  Golden state: F3:F13 filled with IF seniority formulas; G2:G13 filled with concatenation formulas.
  Expected formula pattern for F: =IF(E#>=10,"Senior",IF(E#>=5,"Mid-Level","Junior"))
  Expected formula pattern for G: =B#&" | "&C#&" | $"&TEXT(D#,"0.00")&" | "&F#
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_formula_pattern_concat_004'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Validate sheet exists
    if 'Employees' not in wb.sheetnames:
        print("CRITICAL: 'Employees' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Employees']

    # Precondition gate: verify the sheet has expected data rows (rows 2-13 with employee data)
    data_rows = range(2, 14)  # rows 2 through 13 inclusive
    employees_present = all(ws.cell(row=r, column=1).value is not None for r in data_rows)
    if not employees_present:
        print("CRITICAL: Employee data rows are missing — file structure is corrupted")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Column F (Seniority Level) filled for rows 3-13 (0.5 pts)
    # F2 exists in the initial state; the task requires filling F3:F13 with
    # the same IF formula pattern adjusted per row.
    # -----------------------------------------------------------------------
    try:
        seniority_formula_rows_filled = 0
        seniority_formula_total = 12  # rows 2 through 13 = 12 rows

        # Expected pattern: =IF(E{row}>=10,"Senior",IF(E{row}>=5,"Mid-Level","Junior"))
        # We check rows 3-13 (row 2 already has the formula in initial state)
        for row in range(3, 14):
            val = ws.cell(row=row, column=6).value  # column F
            if val is not None and isinstance(val, str) and val.startswith('='):
                # Check it's an IF formula referencing E{row}
                formula_upper = val.upper().replace(' ', '')
                if (
                    formula_upper.startswith('=IF(') and
                    f'E{row}' in val and
                    'SENIOR' in formula_upper and
                    'MID-LEVEL' in formula_upper and
                    'JUNIOR' in formula_upper
                ):
                    seniority_formula_rows_filled += 1
                    print(f"  PASS F{row}: {val}")
                else:
                    print(f"  FAIL F{row}: formula present but unexpected pattern: {val}")
            else:
                print(f"  FAIL F{row}: empty or not a formula (found: {repr(val)})")

        if seniority_formula_rows_filled == 11:  # rows 3-13 = 11 rows
            print(f"PASS: Component 1 — Column F filled for all rows 3-13 with seniority formula (0.5 pts)")
            total_score += 0.5
        elif seniority_formula_rows_filled >= 6:
            partial = 0.25
            print(f"PARTIAL: Component 1 — Column F filled for {seniority_formula_rows_filled}/11 rows, partial credit ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Column F only filled for {seniority_formula_rows_filled}/11 rows (need rows 3-13)")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Column G (Summary) filled for rows 2-13 with concatenation formula (0.5 pts)
    # Expected pattern: =B{row}&" | "&C{row}&" | $"&TEXT(D{row},"0.00")&" | "&F{row}
    # G column is entirely empty in initial state, so all G2:G13 must be added.
    # -----------------------------------------------------------------------
    try:
        summary_formula_rows_filled = 0

        for row in range(2, 14):
            val = ws.cell(row=row, column=7).value  # column G
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_upper = val.upper().replace(' ', '')
                # Check key elements: concatenation with B, C, TEXT(D,...), F, and the separator " | "
                if (
                    f'B{row}' in val and
                    f'C{row}' in val and
                    f'D{row}' in val and
                    f'F{row}' in val and
                    'TEXT' in formula_upper and
                    '0.00' in val and
                    '|' in val
                ):
                    summary_formula_rows_filled += 1
                    print(f"  PASS G{row}: {val}")
                else:
                    print(f"  FAIL G{row}: formula present but unexpected pattern: {val}")
            else:
                print(f"  FAIL G{row}: empty or not a formula (found: {repr(val)})")

        if summary_formula_rows_filled == 12:  # rows 2-13 = 12 rows
            print(f"PASS: Component 2 — Column G filled for all rows 2-13 with concatenation formula (0.5 pts)")
            total_score += 0.5
        elif summary_formula_rows_filled >= 6:
            partial = 0.25
            print(f"PARTIAL: Component 2 — Column G filled for {summary_formula_rows_filled}/12 rows, partial credit ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Column G only filled for {summary_formula_rows_filled}/12 rows (need rows 2-13)")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM env
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
