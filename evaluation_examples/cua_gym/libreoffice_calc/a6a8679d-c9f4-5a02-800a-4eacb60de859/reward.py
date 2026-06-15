"""
Reward Script: Enable text wrap for cells in column B (B2:B20)
Task ID: calc_fmt_align_wrap_text_032
Domain: libreoffice_calc
Scoring:
  - Component 1: All cells B2:B20 have wrap_text=True          (0.70 pts)
  - Component 2: Column B width unchanged from initial (40.0)  (0.15 pts)
  - Component 3: Header B1, columns A and C not modified       (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_align_wrap_text_032'
SHEET_NAME = 'Task List'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: All cells B2:B20 must have wrap_text=True (0.70 points)
    # This FAILS on initial (all None) → PASSES on golden (all True)
    try:
        cells_with_wrap = 0
        cells_without_wrap = []
        total_cells = 19  # B2 through B20

        for row in range(2, 21):
            cell = ws.cell(row=row, column=2)
            if cell.alignment.wrap_text is True:
                cells_with_wrap += 1
            else:
                cells_without_wrap.append(f"B{row}")

        if cells_with_wrap == total_cells:
            print(f"PASS: Component 1 — All {total_cells} cells B2:B20 have wrap_text=True (0.70 pts)")
            total_score += 0.70
        elif cells_with_wrap > 0:
            # Partial credit: proportional to how many cells are wrapped
            partial = round((cells_with_wrap / total_cells) * 0.70, 4)
            print(f"PARTIAL: Component 1 — {cells_with_wrap}/{total_cells} cells have wrap_text=True "
                  f"({partial:.2f} pts). Missing: {cells_without_wrap[:5]}{'...' if len(cells_without_wrap) > 5 else ''}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No cells in B2:B20 have wrap_text=True. "
                  f"All {total_cells} cells still have wrap_text=None/False")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column B width should remain unchanged (0.15 points)
    # The task states: "column B width should NOT be changed"
    # Initial column B width is 40.0 — verify it hasn't changed in the submitted file
    # This FAILS on initial baseline? No — initial already has 40.0.
    # So we only award this if wrap is applied AND width is preserved.
    # We gate this check: only meaningful to award if component 1 passed (wrap was applied).
    # We check: width is still ~40.0 (within a small tolerance), AND at least some wrap was set.
    try:
        col_b_width = ws.column_dimensions['B'].width
        EXPECTED_WIDTH = 40.0
        TOLERANCE = 0.5

        if cells_with_wrap > 0 and abs((col_b_width or 0) - EXPECTED_WIDTH) <= TOLERANCE:
            print(f"PASS: Component 2 — Column B width preserved at {col_b_width} "
                  f"(expected ~{EXPECTED_WIDTH}) (0.15 pts)")
            total_score += 0.15
        elif cells_with_wrap == 0:
            print(f"SKIP: Component 2 — Skipped because no wrap_text was set (Component 1 failed)")
        else:
            print(f"FAIL: Component 2 — Column B width is {col_b_width}, "
                  f"expected ~{EXPECTED_WIDTH} (should not have been changed)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Non-target cells not modified (0.15 points)
    # B1 (header), A2:A20, C2:C20, D2:D20 should NOT have wrap_text=True
    # This verifies "No other cells should be modified"
    # Initial state: all these are None. Golden should also be None for these.
    # We gate this on component 1: only award if wrap_text was correctly applied to B2:B20
    try:
        non_target_violations = []

        # Check B1 (header should not be wrapped)
        if ws.cell(row=1, column=2).alignment.wrap_text is True:
            non_target_violations.append("B1")

        # Check A2:A20, C2:C20, D2:D20
        for col in [1, 3, 4]:
            for row in range(2, 21):
                cell = ws.cell(row=row, column=col)
                if cell.alignment.wrap_text is True:
                    from openpyxl.utils import get_column_letter
                    non_target_violations.append(f"{get_column_letter(col)}{row}")

        if cells_with_wrap > 0 and len(non_target_violations) == 0:
            print(f"PASS: Component 3 — No non-target cells modified (0.15 pts)")
            total_score += 0.15
        elif cells_with_wrap == 0:
            print(f"SKIP: Component 3 — Skipped because no wrap_text was set (Component 1 failed)")
        else:
            print(f"FAIL: Component 3 — {len(non_target_violations)} non-target cells "
                  f"unexpectedly have wrap_text=True: {non_target_violations[:5]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
