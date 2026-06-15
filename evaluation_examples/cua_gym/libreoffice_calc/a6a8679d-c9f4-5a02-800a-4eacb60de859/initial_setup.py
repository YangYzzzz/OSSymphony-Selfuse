"""
Initial Setup: Enable text wrap for cells in column B (B2:B20)
Task ID: calc_fmt_align_wrap_text_032
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_align_wrap_text_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Task List'

    # Headers
    headers = ['Task ID', 'Description', 'Priority', 'Due Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic task data with long descriptions in column B
    data = [
        ['TASK-001', 'Review and approve the Q3 marketing campaign proposal with budget analysis', 'High', '2025-04-15'],
        ['TASK-002', 'Schedule follow-up meetings with all regional managers before end of month', 'Medium', '2025-04-20'],
        ['TASK-003', 'Prepare the annual compliance report for submission to the regulatory authority', 'High', '2025-04-30'],
        ['TASK-004', 'Coordinate with IT department to upgrade the CRM system to the latest version', 'Medium', '2025-05-05'],
        ['TASK-005', 'Conduct performance reviews for all direct reports in the engineering division', 'High', '2025-05-10'],
        ['TASK-006', 'Finalize the vendor contract negotiations and submit to legal for final review', 'High', '2025-05-12'],
        ['TASK-007', 'Organize the onboarding program for the new batch of summer interns joining in June', 'Low', '2025-05-20'],
        ['TASK-008', 'Analyze customer feedback from the Q2 survey and create actionable improvement plan', 'Medium', '2025-05-25'],
        ['TASK-009', 'Present the product roadmap updates to the executive leadership team for approval', 'High', '2025-06-01'],
        ['TASK-010', 'Update the internal knowledge base with the new operational procedures and guidelines', 'Low', '2025-06-05'],
        ['TASK-011', 'Oversee the migration of legacy data to the new cloud-based data warehouse platform', 'High', '2025-06-10'],
        ['TASK-012', 'Draft the partnership agreement with the new technology provider for digital services', 'Medium', '2025-06-15'],
        ['TASK-013', 'Implement the recommended security patches across all production servers and systems', 'High', '2025-06-18'],
        ['TASK-014', 'Coordinate the cross-departmental budget planning sessions for the upcoming fiscal year', 'Medium', '2025-06-25'],
        ['TASK-015', 'Review and update the disaster recovery plan with input from all department heads', 'Medium', '2025-07-01'],
        ['TASK-016', 'Launch the employee wellness initiative and distribute communication to all staff', 'Low', '2025-07-05'],
        ['TASK-017', 'Complete the market research study on competitor pricing strategies and report findings', 'Medium', '2025-07-10'],
        ['TASK-018', 'Submit the grant application for the research and development innovation fund program', 'High', '2025-07-15'],
        ['TASK-019', 'Plan and execute the company-wide town hall meeting for Q3 strategy alignment', 'Medium', '2025-07-20'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for realistic appearance
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 14

    # IMPORTANT: NO wrap_text on B2:B20 — text overflows (this is the initial state)
    # Explicitly set wrap_text=False to ensure it is disabled
    for row in range(2, 21):
        cell = ws.cell(row=row, column=2)
        cell.alignment = Alignment(wrap_text=False)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
