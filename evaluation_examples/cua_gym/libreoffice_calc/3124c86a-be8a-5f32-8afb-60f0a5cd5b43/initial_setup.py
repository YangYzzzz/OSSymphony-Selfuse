"""
Initial Setup: Enable sheet row grouping by selecting rows 10 through 25,
grouping them, and then collapsing the group to hide those rows.
Task ID: calc_gsi_052
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Budget Report"

    # --- Styling helpers ---
    header_font = Font(name="Arial", size=14, bold=True, color="1F4E79")
    subheader_font = Font(name="Arial", size=11, bold=True, color="2E75B6")
    col_header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="FF2E75B6", end_color="FF2E75B6", fill_type="solid")
    summary_font = Font(name="Arial", size=10, bold=True)
    summary_fill = PatternFill(start_color="FFD6E4F0", end_color="FFD6E4F0", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'

    # --- Row 1: Report title ---
    ws.merge_cells("A1:F1")
    ws["A1"] = "Q1 2025 Departmental Budget Report"
    ws["A1"].font = header_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # --- Row 2: blank ---
    # --- Row 3-4: Summary section ---
    ws["A3"] = "Report Period:"
    ws["B3"] = "January - March 2025"
    ws["A3"].font = subheader_font
    ws["A4"] = "Prepared By:"
    ws["B4"] = "Finance Department"
    ws["A4"].font = subheader_font
    ws["A5"] = "Total Budget:"
    ws["B5"] = 487650.00
    ws["B5"].number_format = currency_fmt
    ws["A5"].font = subheader_font
    ws["A6"] = "Total Spent:"
    ws["B6"] = 412385.75
    ws["B6"].number_format = currency_fmt
    ws["A6"].font = subheader_font
    ws["A7"] = "Variance:"
    ws["B7"] = "=B5-B6"
    ws["B7"].number_format = currency_fmt
    ws["A7"].font = subheader_font

    # --- Row 8: blank separator ---

    # --- Row 9: Column headers for detail section ---
    headers = ["Line Item", "Department", "Category", "Budgeted", "Actual", "Variance"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_idx, value=h)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[9].height = 22

    # --- Rows 10-25: Detail line items (16 rows) ---
    detail_data = [
        ["Office Supplies", "Engineering", "Materials", 12500.00, 11230.50],
        ["Cloud Services", "Engineering", "Technology", 45000.00, 47820.00],
        ["Travel & Conferences", "Engineering", "Travel", 18000.00, 15640.25],
        ["Software Licenses", "Engineering", "Technology", 32000.00, 31500.00],
        ["Recruitment Fees", "HR", "Services", 25000.00, 22100.00],
        ["Training Programs", "HR", "Development", 15000.00, 14750.00],
        ["Benefits Administration", "HR", "Services", 8500.00, 8200.00],
        ["Marketing Campaigns", "Marketing", "Advertising", 65000.00, 68450.00],
        ["Content Creation", "Marketing", "Media", 22000.00, 19800.50],
        ["Event Sponsorships", "Marketing", "Events", 35000.00, 33200.00],
        ["Legal Retainer", "Operations", "Professional", 28000.00, 27500.00],
        ["Facility Maintenance", "Operations", "Facilities", 42000.00, 39875.50],
        ["Insurance Premiums", "Operations", "Insurance", 55000.00, 55000.00],
        ["Equipment Leasing", "Operations", "Equipment", 38000.00, 35120.00],
        ["Consulting Fees", "Finance", "Professional", 30000.00, 28450.00],
        ["Audit & Compliance", "Finance", "Regulatory", 16650.00, 13249.00],
    ]
    for r_offset, row_data in enumerate(detail_data):
        row_num = 10 + r_offset
        ws.cell(row=row_num, column=1, value=row_data[0]).border = thin_border
        ws.cell(row=row_num, column=2, value=row_data[1]).border = thin_border
        ws.cell(row=row_num, column=3, value=row_data[2]).border = thin_border
        c4 = ws.cell(row=row_num, column=4, value=row_data[3])
        c4.number_format = currency_fmt
        c4.border = thin_border
        c5 = ws.cell(row=row_num, column=5, value=row_data[4])
        c5.number_format = currency_fmt
        c5.border = thin_border
        c6 = ws.cell(row=row_num, column=6, value=f"=D{row_num}-E{row_num}")
        c6.number_format = currency_fmt
        c6.border = thin_border

    # --- Row 26: Summary/Total row ---
    ws.cell(row=26, column=1, value="TOTAL").font = summary_font
    ws["A26"].fill = summary_fill
    ws["A26"].border = thin_border
    for col_idx in range(2, 4):
        cell = ws.cell(row=26, column=col_idx)
        cell.fill = summary_fill
        cell.border = thin_border
    ws.cell(row=26, column=4, value="=SUM(D10:D25)").font = summary_font
    ws["D26"].number_format = currency_fmt
    ws["D26"].fill = summary_fill
    ws["D26"].border = thin_border
    ws.cell(row=26, column=5, value="=SUM(E10:E25)").font = summary_font
    ws["E26"].number_format = currency_fmt
    ws["E26"].fill = summary_fill
    ws["E26"].border = thin_border
    ws.cell(row=26, column=6, value="=D26-E26").font = summary_font
    ws["F26"].number_format = currency_fmt
    ws["F26"].fill = summary_fill
    ws["F26"].border = thin_border

    # --- Row 28: Notes ---
    ws["A28"] = "Notes:"
    ws["A28"].font = subheader_font
    ws["A29"] = "All figures are in USD. Variance = Budgeted - Actual."
    ws["A30"] = "Negative variance indicates over-budget spending."

    # --- Column widths ---
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16

    # NO grouping or hidden rows — the task is to add those
    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")


create_initial()
