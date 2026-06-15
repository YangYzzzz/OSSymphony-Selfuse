"""
Initial Setup: Create a workbook with Summary and Sales sheets for INDIRECT formula task.
Task ID: calc_mcp_043
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_043'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sales Sheet ---
    ws_sales = wb.active
    ws_sales.title = 'Sales'

    # Headers
    headers = ['Region', 'Product', 'Units Sold', 'Revenue']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws_sales.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # Sales data (14 rows of data, so row 15 will be a total row)
    sales_data = [
        ['Northeast', 'Widget Pro', 320, 4800],
        ['Southeast', 'Widget Pro', 185, 2775],
        ['Midwest', 'Widget Pro', 410, 6150],
        ['West Coast', 'Widget Pro', 275, 4125],
        ['Northeast', 'Gadget Max', 190, 5700],
        ['Southeast', 'Gadget Max', 145, 4350],
        ['Midwest', 'Gadget Max', 230, 6900],
        ['West Coast', 'Gadget Max', 310, 9300],
        ['Northeast', 'Bolt Standard', 520, 7800],
        ['Southeast', 'Bolt Standard', 340, 5100],
        ['Midwest', 'Bolt Standard', 410, 6150],
        ['West Coast', 'Bolt Standard', 290, 4350],
        ['Northeast', 'Circuit Elite', 155, 6200],
        ['Southeast', 'Circuit Elite', 200, 4800],
    ]

    for r, row_data in enumerate(sales_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_sales.cell(row=r, column=c, value=val)

    # Row 15: Total row with D15 = 78500 (ground truth)
    ws_sales.cell(row=15, column=1, value='TOTAL')
    ws_sales.cell(row=15, column=1).font = Font(bold=True)
    ws_sales.cell(row=15, column=3, value='=SUM(C2:C14)')
    ws_sales.cell(row=15, column=4, value=78500)
    ws_sales.cell(row=15, column=4).font = Font(bold=True)
    ws_sales.cell(row=15, column=4).number_format = '#,##0'

    # Column widths
    ws_sales.column_dimensions['A'].width = 15
    ws_sales.column_dimensions['B'].width = 16
    ws_sales.column_dimensions['C'].width = 12
    ws_sales.column_dimensions['D'].width = 12

    # --- Summary Sheet ---
    ws_summary = wb.create_sheet('Summary')

    # Headers
    ws_summary.cell(row=1, column=1, value='Parameter')
    ws_summary.cell(row=1, column=2, value='Value')
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=12)
    ws_summary.cell(row=1, column=2).font = Font(bold=True, size=12)

    # Row 2: some context
    ws_summary.cell(row=2, column=1, value='Report Date')
    ws_summary.cell(row=2, column=2, value='2025-03-15')

    # Row 3: A3 = 'Sales', B3 = empty (agent must fill with INDIRECT formula)
    ws_summary.cell(row=3, column=1, value='Sales')
    # B3 intentionally left empty

    # Row 4-5: additional context rows
    ws_summary.cell(row=4, column=1, value='Fiscal Quarter')
    ws_summary.cell(row=4, column=2, value='Q1 2025')
    ws_summary.cell(row=5, column=1, value='Prepared By')
    ws_summary.cell(row=5, column=2, value='Elena Rodriguez')

    # Column widths
    ws_summary.column_dimensions['A'].width = 18
    ws_summary.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
