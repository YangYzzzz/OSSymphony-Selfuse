"""
Reward Script: Create INDIRECT formula in Summary!B3 that dynamically
references a cell on another sheet based on the sheet name in A3.
Task ID: calc_mcp_043
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): B3 contains a formula (string starting with '=')
  Component 2 (0.35): Formula uses INDIRECT with a reference to A3
  Component 3 (0.25): Formula targets cell D15 on the referenced sheet
"""

import os
import re
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_043'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet exists
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # Read the raw value of B3
    b3_value = ws['B3'].value
    print(f"DEBUG: Summary!B3 raw value = {repr(b3_value)}")

    # Component 1: B3 contains a formula (0.40 points)
    # In the initial file B3 is empty; in golden it has a formula.
    try:
        if b3_value is not None and isinstance(b3_value, str) and b3_value.strip().startswith('='):
            print(f"PASS: Component 1 — B3 contains a formula: {b3_value} (0.40 pts)")
            total_score += 0.40
        else:
            print(f"FAIL: Component 1 — B3 does not contain a formula. Value: {repr(b3_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula uses INDIRECT with a reference to A3 (0.35 points)
    # The task requires a dynamic reference based on the sheet name in A3,
    # which means the formula must use INDIRECT and reference cell A3.
    try:
        if b3_value is not None and isinstance(b3_value, str):
            formula_upper = b3_value.upper().replace(" ", "")
            has_indirect = 'INDIRECT(' in formula_upper or 'INDIRECT (' in b3_value.upper()
            has_a3_ref = 'A3' in formula_upper
            if has_indirect and has_a3_ref:
                print(f"PASS: Component 2 — Formula uses INDIRECT with A3 reference (0.35 pts)")
                total_score += 0.35
            elif has_indirect:
                print(f"FAIL: Component 2 — Formula uses INDIRECT but does not reference A3")
            elif has_a3_ref:
                print(f"FAIL: Component 2 — Formula references A3 but does not use INDIRECT")
            else:
                print(f"FAIL: Component 2 — Formula lacks both INDIRECT and A3 reference")
        else:
            print(f"FAIL: Component 2 — B3 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula targets cell D15 on the referenced sheet (0.25 points)
    # The formula should construct a reference to D15 on whatever sheet A3 names.
    # Valid patterns include: A3&".D15", A3&"!D15", "'"&A3&"'.D15", etc.
    # We just check that D15 appears in the formula.
    try:
        if b3_value is not None and isinstance(b3_value, str):
            formula_upper = b3_value.upper().replace(" ", "")
            if 'D15' in formula_upper:
                print(f"PASS: Component 3 — Formula targets D15 (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — Formula does not reference D15. Formula: {b3_value}")
        else:
            print(f"FAIL: Component 3 — B3 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
