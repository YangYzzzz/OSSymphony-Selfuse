"""
Initial Setup: Quarterly Revenue Data (4 years, 2020-2023) without YoY % change or 3-year avg rows
Task ID: osworld_calc_annual_pct_change_011
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_annual_pct_change_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Revenue ---
    ws = wb.active
    ws.title = 'Revenue'

    # Header row
    ws['A1'] = 'Year'
    ws['B1'] = 'Q1'
    ws['C1'] = 'Q2'
    ws['D1'] = 'Q3'
    ws['E1'] = 'Q4'

    # Style header row: bold
    header_font = Font(bold=True)
    for col_letter in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col_letter}1'].font = header_font

    # Realistic quarterly revenue data for 4 years (values in thousands)
    # Showing steady growth with seasonal patterns
    data = [
        [2020, 142500, 158300, 175200, 189600],
        [2021, 161800, 179500, 198400, 215700],
        [2022, 183200, 204100, 225300, 248900],
        [2023, 207400, 231600, 256800, 284200],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format year column as text-like integers and revenue columns with number format
    for row in range(2, 6):
        ws.cell(row=row, column=1).number_format = '0'
        for col in range(2, 6):
            ws.cell(row=row, column=col).number_format = '#,##0'

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    for col_letter in ['B', 'C', 'D', 'E']:
        ws.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open LibreOffice Calc with the initial file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
