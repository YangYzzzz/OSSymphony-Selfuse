"""
Reward Script: Add year-over-year percentage change row and rolling 3-year average row
Task ID: osworld_calc_annual_pct_change_011
Domain: libreoffice_calc
Scoring:
  - Component 1: YoY % Change row label present at expected row (0.1 pts)
  - Component 2: YoY % Change formulas in B:E comparing 2023 vs 2022 (0.5 pts)
  - Component 3: Rolling 3-year average row label present below YoY row (0.1 pts)
  - Component 4: AVERAGE formulas in rolling avg row referencing 3 years of data (0.3 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_annual_pct_change_011'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires:
    1. A YoY % change row added after the 4 years of revenue data
       (row 6 in golden, comparing 2023 to 2022 for each quarter B-E)
    2. A rolling 3-year average row added below the YoY row
       (row 7 in golden, AVERAGE of 3 most recent years for each quarter B-E)
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the active/Revenue sheet
    try:
        if 'Revenue' in wb.sheetnames:
            ws = wb['Revenue']
        else:
            ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Pre-condition: verify initial data structure is intact (rows 1-5)
    # If the file doesn't have the expected 4 years of data in rows 2-5, bail.
    # -------------------------------------------------------------------------
    try:
        year_vals = [ws.cell(row=r, column=1).value for r in range(2, 6)]
        if year_vals != [2020, 2021, 2022, 2023]:
            print(f"PRECONDITION FAIL: Expected years [2020,2021,2022,2023] in A2:A5, found {year_vals}")
            print("REWARD: 0.0")
            return 0.0
        else:
            print(f"PRECONDITION PASS: Years in A2:A5 are {year_vals}")
    except Exception as e:
        print(f"PRECONDITION ERROR: Could not verify year labels: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------------
    # Find the YoY % Change row and the 3-Year Average row.
    # They must appear AFTER row 5 (the last data year row).
    # We search rows 6 onward (up to row 10 to be flexible).
    # -------------------------------------------------------------------------
    yoy_row = None
    avg_row = None

    for r in range(6, 15):
        label = ws.cell(row=r, column=1).value
        if label is None:
            continue
        label_str = str(label).strip().lower()
        if yoy_row is None and ('yoy' in label_str or ('change' in label_str and '%' in label_str) or 'pct' in label_str or 'percent' in label_str):
            yoy_row = r
        elif avg_row is None and ('avg' in label_str or 'average' in label_str or '3' in label_str):
            avg_row = r

    # -------------------------------------------------------------------------
    # Component 1: YoY % Change row label present (0.1 pts)
    # -------------------------------------------------------------------------
    try:
        if yoy_row is not None:
            print(f"PASS: Component 1 — YoY % Change label found at row {yoy_row} (A{yoy_row}={repr(ws.cell(row=yoy_row, column=1).value)}) (0.1 pts)")
            total_score += 0.1
        else:
            print("FAIL: Component 1 — No row with YoY/% Change label found after row 5")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: YoY % Change formulas in columns B-E (0.5 pts)
    # The formula should compute (current_year - prev_year) / prev_year
    # i.e., reference the row above this row and two rows above (year 2023 vs 2022)
    # Accept any formula that involves subtraction and division referencing adjacent rows,
    # or that references rows 5 and 4 explicitly.
    # -------------------------------------------------------------------------
    try:
        if yoy_row is None:
            print("FAIL: Component 2 — Cannot check YoY formulas, row not found")
        else:
            formula_cols_pass = 0
            formula_details = []
            for col in range(2, 6):  # B=2, C=3, D=4, E=5
                cell = ws.cell(row=yoy_row, column=col)
                val = cell.value
                col_letter = ['B', 'C', 'D', 'E'][col - 2]

                if val is None:
                    formula_details.append(f"{col_letter}{yoy_row}=None")
                    continue

                val_str = str(val).strip()

                # Check if it's a formula (starts with =)
                if not val_str.startswith('='):
                    formula_details.append(f"{col_letter}{yoy_row}={repr(val_str)} (not a formula)")
                    continue

                # Normalize formula: remove spaces, uppercase
                norm = val_str.upper().replace(' ', '')

                # Check for percentage change pattern:
                # - references row 5 (2023 data) and row 4 (2022 data) explicitly, or
                # - references relative rows that compute a ratio/change
                # Acceptable patterns:
                #   =(B5-B4)/B4   -> references 2023 row vs 2022 row
                #   =(B5/B4)-1    -> equivalent ratio form
                # Also accept if it references yoy_row-1 and yoy_row-2 (relative to yoy_row)
                row_above = yoy_row - 1   # should be row 5 (2023)
                row_above2 = yoy_row - 2  # should be row 4 (2022)

                # Check if formula references the two rows above the yoy_row
                refs_above = (
                    (f'{col_letter}{row_above}' in norm.upper()) and
                    (f'{col_letter}{row_above2}' in norm.upper())
                )
                # Also check for division (percentage change must divide)
                has_division = '/' in norm

                if refs_above and has_division:
                    formula_details.append(f"{col_letter}{yoy_row}={repr(val_str)} OK")
                    formula_cols_pass += 1
                else:
                    # Fallback: check if formula references rows 5 and 4 explicitly
                    # (in case data is not at rows 2-5 but still correct)
                    refs_data_rows = (
                        f'{col_letter}5' in norm and
                        f'{col_letter}4' in norm and
                        has_division
                    )
                    if refs_data_rows:
                        formula_details.append(f"{col_letter}{yoy_row}={repr(val_str)} OK (explicit rows)")
                        formula_cols_pass += 1
                    else:
                        formula_details.append(f"{col_letter}{yoy_row}={repr(val_str)} INVALID (expected division referencing rows {row_above2} and {row_above})")

            print(f"  YoY formula details: {', '.join(formula_details)}")
            if formula_cols_pass == 4:
                print(f"PASS: Component 2 — All 4 YoY % Change formulas correct (0.5 pts)")
                total_score += 0.5
            elif formula_cols_pass >= 2:
                partial = 0.25
                print(f"PARTIAL: Component 2 — {formula_cols_pass}/4 YoY % Change formulas correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 — Only {formula_cols_pass}/4 YoY % Change formulas correct (need >=2 for partial credit)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------------
    # Component 3: Rolling 3-Year Average row label present (0.1 pts)
    # -------------------------------------------------------------------------
    try:
        if avg_row is not None:
            print(f"PASS: Component 3 — 3-Year Average label found at row {avg_row} (A{avg_row}={repr(ws.cell(row=avg_row, column=1).value)}) (0.1 pts)")
            total_score += 0.1
        else:
            print("FAIL: Component 3 — No row with 3-Year Average label found after the YoY row")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------------
    # Component 4: Rolling 3-Year Average formulas in columns B-E (0.3 pts)
    # The formula should compute AVERAGE over 3 years.
    # In the golden file: =AVERAGE(B3:B5) meaning 2021, 2022, 2023 (the 3 most recent years)
    # Accept any AVERAGE formula that spans exactly 3 consecutive rows.
    # -------------------------------------------------------------------------
    try:
        if avg_row is None:
            print("FAIL: Component 4 — Cannot check average formulas, row not found")
        else:
            avg_cols_pass = 0
            avg_details = []
            for col in range(2, 6):  # B=2, C=3, D=4, E=5
                cell = ws.cell(row=avg_row, column=col)
                val = cell.value
                col_letter = ['B', 'C', 'D', 'E'][col - 2]

                if val is None:
                    avg_details.append(f"{col_letter}{avg_row}=None")
                    continue

                val_str = str(val).strip()

                if not val_str.startswith('='):
                    avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} (not a formula)")
                    continue

                norm = val_str.upper().replace(' ', '')

                # Must be an AVERAGE formula
                if 'AVERAGE' not in norm:
                    avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} (no AVERAGE)")
                    continue

                # Extract the range argument from AVERAGE(...)
                # e.g., AVERAGE(B3:B5) → range B3:B5 → 3 rows
                match = re.search(r'AVERAGE\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)', norm)
                if match:
                    start_col_l, start_row_s, end_col_l, end_row_s = match.groups()
                    try:
                        start_row_n = int(start_row_s)
                        end_row_n = int(end_row_s)
                        num_rows = end_row_n - start_row_n + 1

                        # Must span exactly 3 rows and end at the 2023 data row (row 5)
                        # Also: start_col == end_col == col_letter
                        if start_col_l == col_letter and end_col_l == col_letter and num_rows == 3 and end_row_n == 5:
                            avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} OK")
                            avg_cols_pass += 1
                        elif start_col_l == col_letter and end_col_l == col_letter and num_rows == 3:
                            # Accept if span is 3 rows, even if not ending at row 5 (flexible)
                            avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} OK (3-row span)")
                            avg_cols_pass += 1
                        else:
                            avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} INVALID (span={num_rows}, cols={start_col_l}:{end_col_l})")
                    except ValueError:
                        avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} PARSE_ERR")
                else:
                    # Try to parse AVERAGE over non-contiguous refs or other forms
                    # Accept if AVERAGE is present and references the correct column
                    if f'AVERAGE' in norm and col_letter in norm:
                        avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} OK (AVERAGE with col ref)")
                        avg_cols_pass += 1
                    else:
                        avg_details.append(f"{col_letter}{avg_row}={repr(val_str)} UNRECOGNIZED")

            print(f"  Average formula details: {', '.join(avg_details)}")
            if avg_cols_pass == 4:
                print(f"PASS: Component 4 — All 4 rolling 3-year AVERAGE formulas correct (0.3 pts)")
                total_score += 0.3
            elif avg_cols_pass >= 2:
                partial = 0.15
                print(f"PARTIAL: Component 4 — {avg_cols_pass}/4 average formulas correct ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — Only {avg_cols_pass}/4 average formulas correct (need >=2 for partial credit)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
