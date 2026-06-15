"""
Initial Setup: Regional sales workbook with four sheets (North, South, East, West)
Task ID: calc_ps_093
Domain: libreoffice_calc

Creates four sheets with headers and data, no special formatting applied.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_093'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Data for each regional sheet
    regions = {
        'North': [
            ['City', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales'],
            ['Minneapolis', 48230, 51780, 53410, 62150],
            ['Chicago', 87450, 91200, 89670, 95340],
            ['Detroit', 34560, 37890, 36210, 41070],
            ['Milwaukee', 22340, 24560, 23890, 27650],
            ['Cleveland', 31200, 33450, 32780, 36890],
            ['Indianapolis', 41670, 43890, 42350, 47120],
            ['Columbus', 29870, 31450, 30980, 34560],
            ['St. Paul', 18920, 20340, 19780, 22450],
            ['Green Bay', 12450, 13670, 13120, 15340],
            ['Grand Rapids', 16780, 18230, 17560, 19870],
        ],
        'South': [
            ['City', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales'],
            ['Houston', 95670, 98340, 101230, 108450],
            ['Dallas', 82340, 85670, 87890, 93210],
            ['Atlanta', 71230, 74560, 76890, 81340],
            ['Miami', 63450, 66780, 68340, 72560],
            ['Charlotte', 34890, 37120, 38450, 41230],
            ['Nashville', 42560, 44890, 46230, 49670],
            ['New Orleans', 28670, 30450, 31890, 34120],
            ['Tampa', 51230, 53670, 55120, 58340],
            ['Orlando', 47890, 50120, 51670, 54890],
            ['San Antonio', 39450, 41670, 43120, 46230],
        ],
        'East': [
            ['City', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales'],
            ['New York', 142350, 148670, 151230, 162450],
            ['Boston', 78340, 81560, 83890, 89120],
            ['Philadelphia', 65230, 68450, 70120, 74670],
            ['Washington DC', 57890, 60340, 62450, 66780],
            ['Baltimore', 31230, 33450, 34890, 37120],
            ['Pittsburgh', 28670, 30120, 31450, 33890],
            ['Hartford', 19450, 20780, 21560, 23120],
            ['Providence', 15670, 16890, 17450, 18890],
            ['Richmond', 22340, 23890, 24670, 26450],
            ['Newark', 38450, 40120, 41670, 44230],
        ],
        'West': [
            ['City', 'Q1 Sales', 'Q2 Sales', 'Q3 Sales', 'Q4 Sales'],
            ['Los Angeles', 128450, 134670, 137890, 145230],
            ['San Francisco', 96780, 101230, 103450, 109670],
            ['Seattle', 72340, 75890, 77560, 82120],
            ['Denver', 45670, 48120, 49890, 53230],
            ['Phoenix', 54890, 57340, 59120, 62670],
            ['Portland', 38230, 40450, 41890, 44560],
            ['Las Vegas', 42670, 44890, 46340, 49120],
            ['Salt Lake City', 24560, 26120, 27340, 29450],
            ['San Diego', 61230, 64120, 65890, 69340],
            ['Sacramento', 33450, 35120, 36450, 38890],
        ],
    }

    first = True
    for sheet_name, data in regions.items():
        if first:
            ws = wb.active
            ws.title = sheet_name
            first = False
        else:
            ws = wb.create_sheet(sheet_name)

        for r, row_data in enumerate(data, 1):
            for c, val in enumerate(row_data, 1):
                ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
