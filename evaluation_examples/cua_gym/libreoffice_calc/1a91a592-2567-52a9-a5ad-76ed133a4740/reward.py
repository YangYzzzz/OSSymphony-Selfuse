"""
Reward Script: Group-format four regional sheets
Task ID: calc_ps_093
Domain: libreoffice_calc
Scoring:
  Component 1: Bold header row 1 on all 4 sheets (0.4 pts, 0.1 per sheet)
  Component 2: Column A width = 20 on all 4 sheets (0.3 pts, 0.075 per sheet)
  Component 3: Bottom border on row 1 on all 4 sheets (0.3 pts, 0.075 per sheet)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_093'
EXPECTED_SHEETS = ['North', 'South', 'East', 'West']


def persist_app_state(domain: str):
    """Try to save any unsaved changes in LibreOffice."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify all 4 required sheets exist
    for name in EXPECTED_SHEETS:
        if name not in wb.sheetnames:
            print(f"CRITICAL: Sheet '{name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: Bold header row 1 on all 4 sheets (0.4 points, 0.1 per sheet)
    print("\n--- Component 1: Bold headers ---")
    for sheet_name in EXPECTED_SHEETS:
        try:
            ws = wb[sheet_name]
            all_bold = True
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value is not None and not cell.font.bold:
                    all_bold = False
                    break
            if all_bold:
                print(f"  PASS: '{sheet_name}' row 1 is bold (0.1 pts)")
                total_score += 0.1
            else:
                print(f"  FAIL: '{sheet_name}' row 1 is NOT all bold")
        except Exception as e:
            print(f"  ERROR: '{sheet_name}' bold check: {e}")

    # Component 2: Column A width = 20 on all 4 sheets (0.3 points, 0.075 per sheet)
    print("\n--- Component 2: Column A width ---")
    for sheet_name in EXPECTED_SHEETS:
        try:
            ws = wb[sheet_name]
            col_a = ws.column_dimensions.get('A')
            if col_a is not None and col_a.width is not None:
                # Allow small tolerance for floating point
                if abs(col_a.width - 20.0) <= 1.0:
                    print(f"  PASS: '{sheet_name}' column A width = {col_a.width} (0.075 pts)")
                    total_score += 0.075
                else:
                    print(f"  FAIL: '{sheet_name}' column A width = {col_a.width}, expected ~20")
            else:
                print(f"  FAIL: '{sheet_name}' column A width not set (default)")
        except Exception as e:
            print(f"  ERROR: '{sheet_name}' col width check: {e}")

    # Component 3: Bottom border on row 1 across all 4 sheets (0.3 points, 0.075 per sheet)
    print("\n--- Component 3: Bottom border on row 1 ---")
    for sheet_name in EXPECTED_SHEETS:
        try:
            ws = wb[sheet_name]
            all_bordered = True
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value is not None and (cell.border.bottom is None or cell.border.bottom.style is None):
                    all_bordered = False
                    break
            if all_bordered:
                print(f"  PASS: '{sheet_name}' row 1 has bottom border (0.075 pts)")
                total_score += 0.075
            else:
                print(f"  FAIL: '{sheet_name}' row 1 missing bottom border on some cells")
        except Exception as e:
            print(f"  ERROR: '{sheet_name}' border check: {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {final_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
