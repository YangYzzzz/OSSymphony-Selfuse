"""
Initial Setup: Configure Audit Trail sheet with headers/footers and margins
Task ID: calc_mcp_092
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_092'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Audit Trail ---
    ws = wb.active
    ws.title = 'Audit Trail'

    # Headers
    headers = ['Date', 'Entry ID', 'Account', 'Description', 'Debit', 'Credit', 'Balance', 'Auditor']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    white_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Audit data rows - realistic financial audit entries
    data = [
        ['2025-01-05', 'AE-1001', '1010 - Cash', 'Opening balance transfer from prior year', 125000.00, 0.00, 125000.00, 'R. Nakamura'],
        ['2025-01-08', 'AE-1002', '2010 - Accounts Payable', 'Payment to Meridian Supplies Inc.', 0.00, 4350.75, 120649.25, 'S. Okonkwo'],
        ['2025-01-12', 'AE-1003', '4010 - Revenue', 'Invoice #INV-2025-0012 - Consulting services', 18500.00, 0.00, 139149.25, 'R. Nakamura'],
        ['2025-01-15', 'AE-1004', '5020 - Office Supplies', 'Purchase order PO-4488 - Printer cartridges', 0.00, 287.50, 138861.75, 'M. Delacroix'],
        ['2025-01-19', 'AE-1005', '1020 - Petty Cash', 'Petty cash replenishment for Q1', 500.00, 0.00, 139361.75, 'S. Okonkwo'],
        ['2025-01-22', 'AE-1006', '6010 - Utilities', 'Electric bill - January 2025 (Building A)', 0.00, 1245.30, 138116.45, 'M. Delacroix'],
        ['2025-01-25', 'AE-1007', '4010 - Revenue', 'Invoice #INV-2025-0018 - Training workshop fees', 7200.00, 0.00, 145316.45, 'R. Nakamura'],
        ['2025-01-28', 'AE-1008', '5030 - Travel', 'Reimbursement - Client site visit (Chicago)', 0.00, 1892.60, 143423.85, 'S. Okonkwo'],
        ['2025-02-01', 'AE-1009', '2020 - Accrued Liabilities', 'Accrual reversal from December 2024', 3200.00, 0.00, 146623.85, 'M. Delacroix'],
        ['2025-02-05', 'AE-1010', '4020 - Interest Income', 'Bank interest earned - January 2025', 342.18, 0.00, 146966.03, 'R. Nakamura'],
        ['2025-02-08', 'AE-1011', '5010 - Salaries', 'Payroll processing - Pay period 1/16-1/31', 0.00, 28750.00, 118216.03, 'S. Okonkwo'],
        ['2025-02-12', 'AE-1012', '6020 - Insurance', 'Quarterly insurance premium - Property & Liability', 0.00, 4500.00, 113716.03, 'M. Delacroix'],
        ['2025-02-15', 'AE-1013', '4010 - Revenue', 'Invoice #INV-2025-0025 - Software license renewal', 15800.00, 0.00, 129516.03, 'R. Nakamura'],
        ['2025-02-18', 'AE-1014', '5040 - Marketing', 'Digital advertising campaign - February batch', 0.00, 3200.00, 126316.03, 'S. Okonkwo'],
        ['2025-02-22', 'AE-1015', '1010 - Cash', 'Wire transfer received - Client GlobalTech Corp', 42000.00, 0.00, 168316.03, 'M. Delacroix'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c in (5, 6, 7):  # Debit, Credit, Balance columns
                cell.number_format = '#,##0.00'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16

    # NO headers, footers, or margin changes - default state
    # The task is for the agent to add these

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
