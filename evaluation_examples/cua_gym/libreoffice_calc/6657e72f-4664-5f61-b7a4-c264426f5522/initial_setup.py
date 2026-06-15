"""
Initial Setup: Apply conditional formatting to stock quantity column
Task ID: calc_gg2_012
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # --- Inventory Sheet ---
    ws = wb.active
    ws.title = 'Inventory'

    # Headers
    headers = ['Product Code', 'Product Name', 'Category', 'Stock Qty']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Categories and product name prefixes for realistic data
    categories = [
        'Electronics', 'Office Supplies', 'Furniture', 'Cleaning',
        'Safety Equipment', 'Tools', 'Packaging', 'Lighting',
        'Plumbing', 'HVAC'
    ]

    product_prefixes = {
        'Electronics': ['USB Cable', 'Power Adapter', 'Wireless Mouse', 'Keyboard', 'Monitor Stand',
                        'Surge Protector', 'HDMI Cable', 'Webcam', 'Headset', 'Docking Station',
                        'Ethernet Cable', 'Laptop Charger', 'External SSD', 'Bluetooth Speaker',
                        'Smart Plug', 'LED Display', 'Cable Organizer', 'Phone Mount', 'Tablet Case',
                        'Wireless Charger'],
        'Office Supplies': ['Ballpoint Pen', 'Sticky Notes', 'Binder Clip', 'Paper Ream', 'Stapler',
                            'Highlighter Set', 'Desk Organizer', 'File Folder', 'Label Maker', 'Tape Dispenser',
                            'Scissors', 'Whiteboard Marker', 'Paper Clip Box', 'Envelope Pack', 'Notebook',
                            'Index Cards', 'Push Pins', 'Rubber Band Box', 'Correction Tape', 'Pencil Sharpener'],
        'Furniture': ['Ergonomic Chair', 'Standing Desk', 'Bookshelf Unit', 'Filing Cabinet', 'Conference Table',
                      'Desk Lamp', 'Partition Panel', 'Visitor Chair', 'Storage Locker', 'Coat Rack',
                      'Whiteboard', 'Bulletin Board', 'Step Stool', 'Floor Mat', 'Under-Desk Tray',
                      'Cable Management Box', 'Monitor Arm', 'Keyboard Tray', 'Footrest', 'Desk Pad'],
        'Cleaning': ['All-Purpose Cleaner', 'Microfiber Cloth', 'Trash Bag Roll', 'Hand Sanitizer', 'Disinfectant Spray',
                     'Mop Head', 'Broom', 'Dustpan Set', 'Glass Cleaner', 'Floor Wax',
                     'Sponge Pack', 'Paper Towel Roll', 'Air Freshener', 'Latex Gloves Box', 'Bucket',
                     'Squeegee', 'Duster', 'Soap Dispenser', 'Recycling Bin', 'Lint Roller'],
        'Safety Equipment': ['Hard Hat', 'Safety Goggles', 'Hi-Vis Vest', 'Fire Extinguisher', 'First Aid Kit',
                             'Ear Plugs Box', 'Safety Boots', 'Face Shield', 'Respirator Mask', 'Caution Tape',
                             'Emergency Blanket', 'Eye Wash Station', 'Safety Harness', 'Smoke Detector', 'Exit Sign',
                             'Safety Cone', 'Barrier Tape', 'Safety Gloves', 'Knee Pads', 'Back Brace'],
        'Tools': ['Cordless Drill', 'Wrench Set', 'Tape Measure', 'Utility Knife', 'Screwdriver Kit',
                  'Pliers Set', 'Level Tool', 'Hammer', 'Socket Set', 'Wire Stripper',
                  'Allen Key Set', 'Flashlight', 'Clamp Set', 'Hacksaw', 'Voltage Tester',
                  'Heat Gun', 'Caulking Gun', 'Pry Bar', 'Chisel Set', 'File Set'],
        'Packaging': ['Bubble Wrap Roll', 'Shipping Box Small', 'Shipping Box Large', 'Packing Tape', 'Stretch Wrap',
                      'Padded Envelope', 'Packing Peanuts', 'Poly Mailer', 'Cardboard Sheet', 'Edge Protector',
                      'Label Sticker', 'Shrink Wrap', 'Void Fill Paper', 'Strapping Kit', 'Corner Board',
                      'Desiccant Pack', 'Fragile Sticker', 'Zip Lock Bag', 'Foam Insert', 'Tissue Paper'],
        'Lighting': ['LED Bulb 60W', 'Fluorescent Tube', 'Desk Lamp LED', 'Floodlight', 'Emergency Light',
                     'Motion Sensor Light', 'Panel Light', 'Track Light', 'Ceiling Fan Light', 'Under Cabinet Light',
                     'String Lights', 'Spotlight', 'Wall Sconce', 'Pendant Light', 'Recessed Light',
                     'Solar Light', 'Lantern', 'Night Light', 'UV Lamp', 'Work Light'],
        'Plumbing': ['PVC Pipe 1in', 'Pipe Fitting', 'Faucet Washer', 'Teflon Tape', 'Drain Cover',
                     'Water Filter', 'Pipe Clamp', 'Ball Valve', 'Hose Connector', 'Plunger',
                     'Pipe Wrench', 'Sealant Tube', 'Shower Head', 'Toilet Flapper', 'Water Heater Element',
                     'Expansion Tank', 'Pressure Gauge', 'Flow Meter', 'Check Valve', 'Coupling'],
        'HVAC': ['Air Filter 20x25', 'Thermostat', 'Duct Tape', 'Refrigerant Can', 'Blower Motor',
                 'Condensate Pump', 'Insulation Roll', 'Vent Cover', 'Damper', 'Fan Belt',
                 'Compressor Oil', 'Temperature Probe', 'Humidity Sensor', 'Expansion Valve', 'Capacitor',
                 'Contactor', 'Relay Switch', 'Drain Pan', 'UV Germicidal Lamp', 'Coil Cleaner'],
    }

    # Generate 200 product records with varied stock quantities
    # Ensure we have a good mix: some 0, some 1-10, some 11-50, some >50
    stock_distribution = (
        [0] * 15 +                                    # 15 out-of-stock items
        list(range(1, 11)) + [random.randint(1, 10) for _ in range(15)] +  # ~25 critically low
        [random.randint(11, 50) for _ in range(80)] +  # ~80 medium stock
        [random.randint(51, 500) for _ in range(80)]    # ~80 well-stocked
    )
    random.shuffle(stock_distribution)

    row = 2
    product_idx = 0
    for cat in categories:
        products = product_prefixes[cat]
        for prod_name in products:
            if product_idx >= 200:
                break
            code = f'WH-{cat[:3].upper()}-{product_idx + 1001:04d}'
            stock = stock_distribution[product_idx]
            ws.cell(row=row, column=1, value=code)
            ws.cell(row=row, column=2, value=prod_name)
            ws.cell(row=row, column=3, value=cat)
            ws.cell(row=row, column=4, value=stock)
            row += 1
            product_idx += 1

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
