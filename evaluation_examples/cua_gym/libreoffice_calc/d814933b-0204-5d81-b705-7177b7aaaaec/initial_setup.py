"""
Initial Setup: Classroom seating chart with grade integration
Task ID: calc_wf_074
Domain: libreoffice_calc

Creates:
- Roster sheet with 30 students (ID, Name, Grade Average)
- Seating Chart sheet with empty 6x5 grid layout
- Section Analysis sheet with headers only
NO data validation, NO VLOOKUP, NO conditional formatting, NO section formulas.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_074'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1: Roster - 30 students with ID, Name, Grade Average
    # ============================================================
    ws_roster = wb.active
    ws_roster.title = 'Roster'

    headers_roster = ['Student ID', 'Name', 'Grade Average']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    white_font = Font(bold=True, size=11, color="FFFFFF")

    for col, h in enumerate(headers_roster, 1):
        cell = ws_roster.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_align

    students = [
        ('S001', 'Emma Thompson', 94.2),
        ('S002', 'Liam Chen', 87.5),
        ('S003', 'Sophia Martinez', 72.8),
        ('S004', 'Noah Williams', 65.3),
        ('S005', 'Olivia Johnson', 91.0),
        ('S006', 'Ethan Brown', 83.7),
        ('S007', 'Ava Davis', 78.4),
        ('S008', 'Mason Garcia', 56.9),
        ('S009', 'Isabella Rodriguez', 95.1),
        ('S010', 'James Wilson', 88.3),
        ('S011', 'Mia Anderson', 74.6),
        ('S012', 'Benjamin Lee', 69.2),
        ('S013', 'Charlotte Taylor', 92.7),
        ('S014', 'Alexander Moore', 81.4),
        ('S015', 'Amelia Jackson', 77.0),
        ('S016', 'Daniel White', 63.8),
        ('S017', 'Harper Harris', 90.5),
        ('S018', 'Michael Clark', 85.2),
        ('S019', 'Evelyn Lewis', 71.3),
        ('S020', 'Sebastian Walker', 58.7),
        ('S021', 'Abigail Hall', 93.6),
        ('S022', 'Jack Allen', 82.1),
        ('S023', 'Emily Young', 76.9),
        ('S024', 'Owen King', 67.4),
        ('S025', 'Elizabeth Wright', 96.3),
        ('S026', 'Lucas Scott', 84.8),
        ('S027', 'Sofia Green', 73.5),
        ('S028', 'Henry Adams', 61.2),
        ('S029', 'Aria Nelson', 89.7),
        ('S030', 'William Baker', 79.6),
    ]

    for r, (sid, name, grade) in enumerate(students, 2):
        ws_roster.cell(row=r, column=1, value=sid)
        ws_roster.cell(row=r, column=2, value=name)
        ws_roster.cell(row=r, column=3, value=grade)
        ws_roster.cell(row=r, column=3).number_format = '0.0'

    ws_roster.column_dimensions['A'].width = 12
    ws_roster.column_dimensions['B'].width = 22
    ws_roster.column_dimensions['C'].width = 16

    # ============================================================
    # Sheet 2: Seating Chart - 6 rows x 5 columns grid layout
    # ============================================================
    ws_seating = wb.create_sheet('Seating Chart')

    # Title row
    ws_seating.merge_cells('A1:J1')
    title_cell = ws_seating.cell(row=1, column=1, value='Classroom Seating Chart')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers for desk columns (A through E)
    desk_labels = ['Seat A', 'Seat B', 'Seat C', 'Seat D', 'Seat E']
    # Layout: Each seating row uses 2 spreadsheet rows (name + grade label)
    # Row 2: blank spacer
    # Row 3: "Row 1" label + desk headers
    ws_seating.cell(row=3, column=1, value='').font = Font(bold=True)
    for col_idx, label in enumerate(desk_labels, 2):
        cell = ws_seating.cell(row=2, column=col_idx, value=label)
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Build the grid: 6 seating rows, each taking 2 spreadsheet rows
    # Row labels in column A, student name cells in columns B-F
    # Grade display row beneath each name row
    for seat_row in range(1, 7):
        name_excel_row = 3 + (seat_row - 1) * 2  # rows 3,5,7,9,11,13
        grade_excel_row = name_excel_row + 1       # rows 4,6,8,10,12,14

        # Row label
        ws_seating.cell(row=name_excel_row, column=1, value=f'Row {seat_row}')
        ws_seating.cell(row=name_excel_row, column=1).font = Font(bold=True, size=10)
        ws_seating.cell(row=grade_excel_row, column=1, value='Grade:')
        ws_seating.cell(row=grade_excel_row, column=1).font = Font(italic=True, size=9, color="666666")

        for desk_col in range(2, 7):  # columns B through F
            # Name cell - empty, with border
            name_cell = ws_seating.cell(row=name_excel_row, column=desk_col, value='')
            name_cell.border = thin_border
            name_cell.alignment = Alignment(horizontal="center", vertical="center")

            # Grade cell - empty, with border
            grade_cell = ws_seating.cell(row=grade_excel_row, column=desk_col, value='')
            grade_cell.border = thin_border
            grade_cell.alignment = Alignment(horizontal="center", vertical="center")
            grade_cell.number_format = '0.0'

    # Set column widths
    ws_seating.column_dimensions['A'].width = 10
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_seating.column_dimensions[col_letter].width = 22

    ws_seating.row_dimensions[1].height = 28

    # ============================================================
    # Sheet 3: Section Analysis - headers only, no formulas
    # ============================================================
    ws_analysis = wb.create_sheet('Section Analysis')

    ws_analysis.merge_cells('A1:D1')
    analysis_title = ws_analysis.cell(row=1, column=1, value='Section Performance Analysis')
    analysis_title.font = Font(bold=True, size=13)
    analysis_title.alignment = Alignment(horizontal="center")

    analysis_headers = ['Section', 'Rows', 'Average Grade', 'Performance Level']
    for col, h in enumerate(analysis_headers, 1):
        cell = ws_analysis.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Section labels only - no formulas
    sections = [
        ('Front', 'Rows 1-2'),
        ('Middle', 'Rows 3-4'),
        ('Back', 'Rows 5-6'),
    ]
    for r, (section, rows) in enumerate(sections, 4):
        ws_analysis.cell(row=r, column=1, value=section)
        ws_analysis.cell(row=r, column=2, value=rows)
        # Columns C and D are intentionally empty - task is to fill them

    # Summary row label
    ws_analysis.cell(row=8, column=1, value='Overall Summary')
    ws_analysis.cell(row=8, column=1).font = Font(bold=True)

    ws_analysis.column_dimensions['A'].width = 14
    ws_analysis.column_dimensions['B'].width = 12
    ws_analysis.column_dimensions['C'].width = 16
    ws_analysis.column_dimensions['D'].width = 20

    # Save
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
