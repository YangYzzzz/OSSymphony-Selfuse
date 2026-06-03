"""
Reward Script: Classroom Seating Chart with Grade Integration
Task ID: calc_wf_074
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.25): Student names populated in 6x5 seating grid
  - Component 2 (0.25): VLOOKUP formulas in grade rows linking to Roster
  - Component 3 (0.15): Data validation (dropdown) on name cells from Roster
  - Component 4 (0.15): Conditional formatting rules on grade rows (4 color rules)
  - Component 5 (0.20): Section Analysis formulas (AVERAGE, IF, summary)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_074'

# Known roster names from Roster sheet B2:B31
ROSTER_NAMES = [
    'Emma Thompson', 'Liam Chen', 'Sophia Martinez', 'Noah Williams',
    'Olivia Johnson', 'Ethan Brown', 'Ava Davis', 'Mason Garcia',
    'Isabella Rodriguez', 'James Wilson', 'Mia Anderson', 'Benjamin Lee',
    'Charlotte Taylor', 'Alexander Moore', 'Amelia Jackson', 'Daniel White',
    'Harper Harris', 'Michael Clark', 'Evelyn Lewis', 'Sebastian Walker',
    'Abigail Hall', 'Jack Allen', 'Emily Young', 'Owen King',
    'Elizabeth Wright', 'Lucas Scott', 'Sofia Green', 'Henry Adams',
    'Aria Nelson', 'William Baker',
]

# Seating grid: name rows are 3,5,7,9,11,13; grade rows are 4,6,8,10,12,14
# Columns B-F (2-6)
NAME_ROWS = [3, 5, 7, 9, 11, 13]
GRADE_ROWS = [4, 6, 8, 10, 12, 14]
SEAT_COLS = [2, 3, 4, 5, 6]  # B through F


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: required sheets exist
    required_sheets = ['Roster', 'Seating Chart', 'Section Analysis']
    for sn in required_sheets:
        if sn not in wb.sheetnames:
            print(f"CRITICAL: Missing required sheet '{sn}'")
            print("REWARD: 0.0")
            return 0.0

    ws_seat = wb['Seating Chart']
    ws_analysis = wb['Section Analysis']

    # =========================================================================
    # Component 1: Student names populated in seating grid (0.25 points)
    # In initial_env, all name cells (B3:F3, B5:F5, ..., B13:F13) are None.
    # In golden_env, they should contain valid student names from the roster.
    # =========================================================================
    try:
        filled_name_count = 0
        total_name_cells = len(NAME_ROWS) * len(SEAT_COLS)  # 30 cells
        for r in NAME_ROWS:
            for c in SEAT_COLS:
                val = ws_seat.cell(row=r, column=c).value
                if val is not None and str(val).strip() in ROSTER_NAMES:
                    filled_name_count += 1

        if filled_name_count >= 25:
            # At least 25 out of 30 seats filled with valid roster names
            print(f"PASS: Component 1 - {filled_name_count}/{total_name_cells} seats filled with roster names (0.25 pts)")
            total_score += 0.25
        elif filled_name_count >= 15:
            partial = 0.15
            print(f"PARTIAL: Component 1 - {filled_name_count}/{total_name_cells} seats filled ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {filled_name_count}/{total_name_cells} seats filled with roster names")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: VLOOKUP formulas in grade rows (0.25 points)
    # In initial_env, grade cells (B4:F4, B6:F6, ..., B14:F14) are None.
    # In golden_env, they should contain VLOOKUP formulas referencing Roster.
    # =========================================================================
    try:
        vlookup_count = 0
        total_grade_cells = len(GRADE_ROWS) * len(SEAT_COLS)  # 30 cells
        for r in GRADE_ROWS:
            for c in SEAT_COLS:
                val = ws_seat.cell(row=r, column=c).value
                if val is not None and isinstance(val, str) and 'VLOOKUP' in val.upper():
                    vlookup_count += 1

        if vlookup_count >= 25:
            print(f"PASS: Component 2 - {vlookup_count}/{total_grade_cells} grade cells have VLOOKUP formulas (0.25 pts)")
            total_score += 0.25
        elif vlookup_count >= 15:
            partial = 0.15
            print(f"PARTIAL: Component 2 - {vlookup_count}/{total_grade_cells} VLOOKUP formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {vlookup_count}/{total_grade_cells} grade cells have VLOOKUP formulas")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Data validation on name cells (0.15 points)
    # In initial_env, there are no data validations.
    # In golden_env, name cells should have list-type validation from Roster.
    # =========================================================================
    try:
        has_dv = False
        dv_list = ws_seat.data_validations.dataValidation if ws_seat.data_validations else []
        for dv in dv_list:
            if dv.type == 'list':
                # Check that it references Roster names
                formula_str = str(dv.formula1) if dv.formula1 else ''
                if 'Roster' in formula_str or 'roster' in formula_str.lower():
                    # Check that it applies to at least some name cells
                    sqref_str = str(dv.sqref) if dv.sqref else ''
                    # Name cells are in B,C,D,E,F at rows 3,5,7,9,11,13
                    name_cell_refs = []
                    for r in NAME_ROWS:
                        for c in SEAT_COLS:
                            cell_ref = openpyxl.utils.get_column_letter(c) + str(r)
                            name_cell_refs.append(cell_ref)
                    matched_cells = sum(1 for ref in name_cell_refs if ref in sqref_str)
                    if matched_cells >= 20:
                        has_dv = True
                        break

        if has_dv:
            print(f"PASS: Component 3 - Data validation (list from Roster) found on name cells (0.15 pts)")
            total_score += 0.15
        else:
            # Also accept any list-type DV on name cells even if formula doesn't mention Roster by name
            for dv in dv_list:
                if dv.type == 'list':
                    sqref_str = str(dv.sqref) if dv.sqref else ''
                    name_cell_refs = []
                    for r in NAME_ROWS:
                        for c in SEAT_COLS:
                            cell_ref = openpyxl.utils.get_column_letter(c) + str(r)
                            name_cell_refs.append(cell_ref)
                    matched_cells = sum(1 for ref in name_cell_refs if ref in sqref_str)
                    if matched_cells >= 10:
                        has_dv = True
                        print(f"PASS: Component 3 - Data validation (list) found on {matched_cells} name cells (0.15 pts)")
                        total_score += 0.15
                        break

            if not has_dv:
                print(f"FAIL: Component 3 - No list-type data validation found on seating name cells")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =========================================================================
    # Component 4: Conditional formatting on grade rows (0.15 points)
    # In initial_env, there is no conditional formatting.
    # In golden_env, grade rows should have color rules (green/blue/yellow/red).
    # =========================================================================
    try:
        cf_rules = list(ws_seat.conditional_formatting)
        # Count how many grade row ranges have conditional formatting
        grade_row_ranges_with_cf = 0
        total_cf_rules = 0
        for cf in cf_rules:
            range_str = str(cf)
            # Check if this CF range covers any grade row cells
            for r in GRADE_ROWS:
                row_ref = str(r)
                # Check if the range includes cells from this grade row
                if f'B{r}' in range_str or f'{r}' in range_str:
                    grade_row_ranges_with_cf += 1
                    total_cf_rules += len(cf.rules)
                    break

        if grade_row_ranges_with_cf >= 5 and total_cf_rules >= 16:
            print(f"PASS: Component 4 - Conditional formatting on {grade_row_ranges_with_cf} grade rows, {total_cf_rules} total rules (0.15 pts)")
            total_score += 0.15
        elif grade_row_ranges_with_cf >= 3 and total_cf_rules >= 8:
            partial = 0.08
            print(f"PARTIAL: Component 4 - CF on {grade_row_ranges_with_cf} rows, {total_cf_rules} rules ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - Only {grade_row_ranges_with_cf} grade rows with CF, {total_cf_rules} total rules")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =========================================================================
    # Component 5: Section Analysis formulas (0.20 points)
    # In initial_env, C4:C6 and D4:D6 are None, no summary rows.
    # In golden_env, they should have AVERAGE formulas, IF performance levels,
    # and summary stats (overall, highest, lowest, spread).
    # =========================================================================
    try:
        analysis_score = 0.0

        # Sub-check 5a: AVERAGE formulas in C4:C6 (0.10 pts)
        avg_count = 0
        for r in [4, 5, 6]:
            val = ws_analysis.cell(row=r, column=3).value  # Column C
            if val is not None and isinstance(val, str) and 'AVERAGE' in val.upper():
                avg_count += 1
        if avg_count >= 3:
            analysis_score += 0.10
            print(f"PASS: Component 5a - AVERAGE formulas in C4:C6 ({avg_count}/3) (0.10 pts)")
        else:
            print(f"FAIL: Component 5a - Only {avg_count}/3 section cells have AVERAGE formulas")

        # Sub-check 5b: IF formulas for performance levels in D4:D6 (0.05 pts)
        if_count = 0
        for r in [4, 5, 6]:
            val = ws_analysis.cell(row=r, column=4).value  # Column D
            if val is not None and isinstance(val, str) and 'IF' in val.upper():
                if_count += 1
        if if_count >= 3:
            analysis_score += 0.05
            print(f"PASS: Component 5b - IF formulas in D4:D6 ({if_count}/3) (0.05 pts)")
        else:
            print(f"FAIL: Component 5b - Only {if_count}/3 section cells have IF performance formulas")

        # Sub-check 5c: Summary rows (C8 overall, or additional stats) (0.05 pts)
        summary_formulas = 0
        # Check C8 for overall average
        c8 = ws_analysis.cell(row=8, column=3).value
        if c8 is not None and isinstance(c8, str) and ('AVERAGE' in c8.upper() or 'SUM' in c8.upper()):
            summary_formulas += 1
        # Check for additional summary rows (MAX, MIN, etc.)
        for r in range(9, ws_analysis.max_row + 1):
            for c in [2, 3]:
                val = ws_analysis.cell(row=r, column=c).value
                if val is not None and isinstance(val, str) and any(f in val.upper() for f in ['MAX', 'MIN', 'INDEX', 'MATCH']):
                    summary_formulas += 1
        if summary_formulas >= 2:
            analysis_score += 0.05
            print(f"PASS: Component 5c - Summary formulas found ({summary_formulas}) (0.05 pts)")
        else:
            print(f"FAIL: Component 5c - Only {summary_formulas} summary formulas found")

        total_score += analysis_score

    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
