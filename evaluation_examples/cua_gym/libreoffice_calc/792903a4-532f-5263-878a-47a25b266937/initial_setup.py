"""
Initial Setup: Create a spreadsheet with score data for color scale macro task
Task ID: calc_mcp_027
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_027'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

# Deterministic score values covering all three ranges (0-33, 34-66, 67-100)
SCORES = [
    12, 45, 78, 5, 91, 33, 67, 50, 22, 88,
    15, 70, 40, 3, 99, 28, 55, 82, 61, 8,
    73, 36, 19, 94, 47, 0, 100, 64, 11, 85,
    27, 58, 76, 42, 17, 90, 34, 66, 23, 71,
    53, 9, 87, 31, 60, 96, 14, 48, 79, 0,  # extra to ensure 49 values
]
# Use exactly 49 values for B2:B50
SCORES = SCORES[:49]


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Scores'

    # Headers
    ws.cell(row=1, column=1, value='ID')
    ws.cell(row=1, column=2, value='Score')

    # Data rows B2:B50 (rows 2 through 50 = 49 data rows)
    for i, score in enumerate(SCORES):
        row = i + 2
        ws.cell(row=row, column=1, value=i + 1)  # ID
        ws.cell(row=row, column=2, value=score)   # Score (integer 0-100)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
