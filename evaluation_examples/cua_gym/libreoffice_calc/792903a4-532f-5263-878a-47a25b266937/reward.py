"""
Reward Script: Color Scale Macro Verification
Task ID: calc_mcp_027
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Red cells (values 0-33) have FFFF0000 background
  Component 2 (0.3): Yellow cells (values 34-66) have FFFFFF00 background
  Component 3 (0.3): Green cells (values 67-100) have FF00FF00 background
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_027'

# Expected color mappings (ARGB format)
RED = 'FFFF0000'
YELLOW = 'FFFFFF00'
GREEN = 'FF00FF00'


def get_fill_rgb(cell):
    """Safely extract the fill foreground color RGB string."""
    try:
        if cell.fill.fill_type == 'solid' and cell.fill.fgColor and cell.fill.fgColor.rgb:
            return str(cell.fill.fgColor.rgb)
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Scores' sheet exists
    if 'Scores' not in wb.sheetnames:
        print("FAIL: 'Scores' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Scores']

    # Classify cells by value range
    red_cells = []     # 0-33
    yellow_cells = []  # 34-66
    green_cells = []   # 67-100

    for r in range(2, 51):
        cell = ws.cell(row=r, column=2)
        val = cell.value
        if val is None:
            continue
        try:
            v = int(val)
        except (ValueError, TypeError):
            continue

        if 0 <= v <= 33:
            red_cells.append((r, v))
        elif 34 <= v <= 66:
            yellow_cells.append((r, v))
        elif 67 <= v <= 100:
            green_cells.append((r, v))

    # Component 1: Red background for values 0-33 (0.4 points)
    try:
        if len(red_cells) == 0:
            print("FAIL: Component 1 -- No cells in 0-33 range found")
        else:
            correct = 0
            for r, v in red_cells:
                rgb = get_fill_rgb(ws.cell(row=r, column=2))
                if rgb == RED:
                    correct += 1
                else:
                    print(f"  DETAIL: B{r} (value={v}) expected {RED}, got {rgb}")
            ratio = correct / len(red_cells)
            if ratio == 1.0:
                print(f"PASS: Component 1 -- All {len(red_cells)} red-range cells have red fill (0.4 pts)")
                total_score += 0.4
            elif ratio > 0:
                pts = round(0.4 * ratio, 4)
                print(f"PARTIAL: Component 1 -- {correct}/{len(red_cells)} red-range cells correct ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 1 -- 0/{len(red_cells)} red-range cells have red fill")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Yellow background for values 34-66 (0.3 points)
    try:
        if len(yellow_cells) == 0:
            print("FAIL: Component 2 -- No cells in 34-66 range found")
        else:
            correct = 0
            for r, v in yellow_cells:
                rgb = get_fill_rgb(ws.cell(row=r, column=2))
                if rgb == YELLOW:
                    correct += 1
                else:
                    print(f"  DETAIL: B{r} (value={v}) expected {YELLOW}, got {rgb}")
            ratio = correct / len(yellow_cells)
            if ratio == 1.0:
                print(f"PASS: Component 2 -- All {len(yellow_cells)} yellow-range cells have yellow fill (0.3 pts)")
                total_score += 0.3
            elif ratio > 0:
                pts = round(0.3 * ratio, 4)
                print(f"PARTIAL: Component 2 -- {correct}/{len(yellow_cells)} yellow-range cells correct ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 2 -- 0/{len(yellow_cells)} yellow-range cells have yellow fill")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Green background for values 67-100 (0.3 points)
    try:
        if len(green_cells) == 0:
            print("FAIL: Component 3 -- No cells in 67-100 range found")
        else:
            correct = 0
            for r, v in green_cells:
                rgb = get_fill_rgb(ws.cell(row=r, column=2))
                if rgb == GREEN:
                    correct += 1
                else:
                    print(f"  DETAIL: B{r} (value={v}) expected {GREEN}, got {rgb}")
            ratio = correct / len(green_cells)
            if ratio == 1.0:
                print(f"PASS: Component 3 -- All {len(green_cells)} green-range cells have green fill (0.3 pts)")
                total_score += 0.3
            elif ratio > 0:
                pts = round(0.3 * ratio, 4)
                print(f"PARTIAL: Component 3 -- {correct}/{len(green_cells)} green-range cells correct ({pts} pts)")
                total_score += pts
            else:
                print(f"FAIL: Component 3 -- 0/{len(green_cells)} green-range cells have green fill")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
