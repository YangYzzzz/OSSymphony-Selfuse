"""
Initial Setup: Grade sheet with student assignments (no freeze panes)
Task ID: calc_edu_freeze_student_names_008
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_freeze_student_names_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Assignments'

    # --- Headers ---
    # Column A: Student ID, Column B: Student Name, Columns C-AD: Assignment 1-28
    headers = ['Student ID', 'Student Name']
    for i in range(1, 29):
        headers.append(f'Assignment {i}')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Student Data ---
    # 60 realistic students
    first_names = [
        'Emma', 'Liam', 'Olivia', 'Noah', 'Ava', 'William', 'Isabella', 'James',
        'Sophia', 'Oliver', 'Mia', 'Benjamin', 'Charlotte', 'Elijah', 'Amelia',
        'Lucas', 'Harper', 'Mason', 'Evelyn', 'Aiden', 'Abigail', 'Logan',
        'Emily', 'Ethan', 'Elizabeth', 'Jackson', 'Mila', 'Sebastian', 'Ella',
        'Mateo', 'Avery', 'Jack', 'Sofia', 'Owen', 'Camila', 'Theodore',
        'Aria', 'Aiden', 'Scarlett', 'Henry', 'Victoria', 'Samuel', 'Madison',
        'Alexander', 'Luna', 'Michael', 'Grace', 'Daniel', 'Chloe', 'Matthew',
        'Penelope', 'Jacob', 'Layla', 'Logan', 'Riley', 'Ryan', 'Zoey',
        'Nathan', 'Nora', 'David', 'Lily'
    ]

    last_names = [
        'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller', 'Davis',
        'Martinez', 'Hernandez', 'Wilson', 'Anderson', 'Taylor', 'Thomas',
        'Jackson', 'White', 'Harris', 'Martin', 'Thompson', 'Moore', 'Young',
        'Allen', 'King', 'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill',
        'Flores', 'Green', 'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera',
        'Campbell', 'Mitchell', 'Carter', 'Roberts', 'Gomez', 'Phillips',
        'Evans', 'Turner', 'Diaz', 'Parker', 'Cruz', 'Edwards', 'Collins',
        'Reyes', 'Stewart', 'Morris', 'Morales', 'Murphy', 'Cook', 'Rogers',
        'Gutierrez', 'Ortiz', 'Morgan', 'Cooper', 'Peterson', 'Bailey'
    ]

    # Scores - varied realistic assignment grades
    import random
    random.seed(42)

    for i in range(60):
        student_id = f'STU{1001 + i}'
        student_name = f'{first_names[i]} {last_names[i]}'
        row_data = [student_id, student_name]

        # Generate 28 assignment scores (out of 100)
        for _ in range(28):
            # Realistic grade distribution: most students score 60-100
            score = random.randint(55, 100)
            row_data.append(score)

        for col, val in enumerate(row_data, 1):
            ws.cell(row=i + 2, column=col, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 20
    for col_letter in 'CDEFGHIJKLMNOPQRSTUVWXYZ':
        ws.column_dimensions[col_letter].width = 13
    ws.column_dimensions['AA'].width = 13
    ws.column_dimensions['AB'].width = 13
    ws.column_dimensions['AC'].width = 13
    ws.column_dimensions['AD'].width = 13

    # Explicitly ensure NO freeze panes are set
    ws.freeze_panes = None

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Assignments')
    print(f'  Rows: 61 (1 header + 60 students)')
    print(f'  Columns: 30 (Student ID, Student Name, Assignment 1-28)')
    print(f'  Freeze panes: None')


create_initial()
