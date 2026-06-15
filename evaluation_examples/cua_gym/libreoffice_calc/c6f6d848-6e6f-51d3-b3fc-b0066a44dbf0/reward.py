"""
Reward Script: Freeze first two columns and header row in grade sheet
Task ID: calc_edu_freeze_student_names_008
Domain: libreoffice_calc
Scoring:
  - Component 1: Freeze panes are set (not None) in 'Assignments' sheet — 0.4 pts
  - Component 2: Freeze point is exactly 'C2' (freezes row 1 + columns A & B) — 0.6 pts
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_freeze_student_names_008'
SHEET_NAME = 'Assignments'
EXPECTED_FREEZE = 'C2'


def verify_task(file_path):
    """
    Verify that freeze panes are correctly set to C2 in the Assignments sheet.
    Freeze at C2 means: row 1 (header) stays visible when scrolling down,
    and columns A and B (Student ID and Student Name) stay visible when scrolling right.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: sheet 'Assignments' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Available sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Freeze panes are set (not None) — 0.4 points
    # This FAILS on initial (freeze_panes=None), PASSES on golden (freeze_panes='C2')
    try:
        fp = ws.freeze_panes
        if fp is not None and fp != '':
            print(f"PASS: Component 1 — Freeze panes are set (value: {repr(fp)}) (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Freeze panes not set (found: {repr(fp)}). Expected a non-None value.")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not read freeze_panes: {e}")

    # Component 2: Freeze point is exactly 'C2' — 0.6 points
    # C2 means: freeze after row 1 (header) AND after column B (Student ID + Student Name)
    # This FAILS on initial (freeze_panes=None), PASSES on golden (freeze_panes='C2')
    try:
        fp = ws.freeze_panes
        if fp == EXPECTED_FREEZE:
            print(f"PASS: Component 2 — Freeze point is exactly '{EXPECTED_FREEZE}' (freezes row 1 + columns A & B) (0.6 pts)")
            total_score += 0.6
        else:
            print(f"FAIL: Component 2 — Expected freeze_panes='{EXPECTED_FREEZE}', found: {repr(fp)}")
            if fp is not None and fp != '':
                # Partial diagnosis
                try:
                    from openpyxl.utils.cell import coordinate_to_tuple
                    row, col = coordinate_to_tuple(fp)
                    if row == 2 and col != 3:
                        print(f"  Note: Row freeze is correct (row=2), but column freeze is wrong (col={col}, expected 3)")
                    elif row != 2 and col == 3:
                        print(f"  Note: Column freeze is correct (col=3), but row freeze is wrong (row={row}, expected 2)")
                except Exception:
                    pass
    except Exception as e:
        print(f"ERROR: Component 2 — Could not verify freeze point: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
