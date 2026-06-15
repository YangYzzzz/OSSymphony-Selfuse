"""
Reward Script: Create automated narrative sentences in column E using TEXT() formulas.
Task ID: osworld_calc_text_format_number_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.2): At least one narrative formula present in column E (rows 2+)
  Component 2 (0.4): All 15 transaction rows (E2:E16) contain formulas
  Component 3 (0.4): Formulas use correct TEXT() date format DD-MMM-YYYY and amount format #,##0.00
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_text_format_number_008'

# Expected formula structure patterns
# The formula should embed:
#   TEXT(<date_cell>, "DD-MMM-YYYY")  for date
#   TEXT(<amount_cell>, "#,##0.00")   for amount
# And must reference the corresponding row's B and D columns

DATE_FORMAT_PATTERN = re.compile(r'TEXT\s*\(\s*B\d+\s*,\s*"DD-MMM-YYYY"\s*\)', re.IGNORECASE)
AMOUNT_FORMAT_PATTERN = re.compile(r'TEXT\s*\(\s*D\d+\s*,\s*"#,##0\.00"\s*\)', re.IGNORECASE)


def verify_task(file_path):
    """
    Verify task completion: column E should contain narrative text formulas
    that embed date (DD-MMM-YYYY), category, and amount (#,##0.00) for each transaction row.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Get the active sheet (expected: 'Transactions')
    try:
        ws = wb.active
        print(f"INFO: Active sheet = '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: E1 header should be 'Narrative' (sanity check that this is the right file)
    try:
        e1_val = ws.cell(row=1, column=5).value
        if e1_val is None or str(e1_val).strip().lower() != 'narrative':
            print(f"WARN: E1 header is '{e1_val}', expected 'Narrative'. Proceeding anyway.")
    except Exception as e:
        print(f"WARN: Could not read E1 header: {e}")

    # --- Component 1: At least one narrative formula present in column E (rows 2+) ---
    # (0.2 points) — checks that the task was started at all
    try:
        formulas_found = []
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=5).value
            if cell_val is not None:
                formulas_found.append((row, cell_val))

        if len(formulas_found) >= 1:
            print(f"PASS: Component 1 — {len(formulas_found)} formula(s) found in column E (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 1 — No formulas found in column E rows 2+")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # --- Component 2: All 15 transaction rows (E2:E16) contain formulas ---
    # (0.4 points) — checks completeness across all transaction rows
    try:
        # We expect rows 2..16 (15 data rows)
        expected_rows = list(range(2, 17))  # rows 2-16 inclusive
        missing_rows = []
        filled_rows = []

        for row in expected_rows:
            cell_val = ws.cell(row=row, column=5).value
            if cell_val is not None:
                filled_rows.append(row)
            else:
                missing_rows.append(row)

        if len(missing_rows) == 0:
            print(f"PASS: Component 2 — All 15 rows (E2:E16) contain narrative formulas (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 2 — {len(filled_rows)}/15 rows filled. Missing rows: {missing_rows}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # --- Component 3: Formulas use correct TEXT() date and amount formats ---
    # (0.4 points) — checks that the formula structure matches the required pattern:
    #   TEXT(<date>, "DD-MMM-YYYY") and TEXT(<amount>, "#,##0.00")
    # Both patterns must appear in each formula for full credit.
    # Partial scoring: 0.2 for date format correct, 0.2 for amount format correct
    try:
        date_format_correct_count = 0
        amount_format_correct_count = 0
        formula_count = 0

        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=5).value
            if cell_val is None:
                continue
            formula_str = str(cell_val)
            formula_count += 1

            # Check date format: TEXT(B<row>, "DD-MMM-YYYY")
            if DATE_FORMAT_PATTERN.search(formula_str):
                date_format_correct_count += 1

            # Check amount format: TEXT(D<row>, "#,##0.00")
            if AMOUNT_FORMAT_PATTERN.search(formula_str):
                amount_format_correct_count += 1

        if formula_count == 0:
            print("FAIL: Component 3 — No formulas to evaluate")
        else:
            date_ok = date_format_correct_count == formula_count
            amount_ok = amount_format_correct_count == formula_count

            if date_ok:
                print(f"PASS: Component 3a — All {formula_count} formulas use DD-MMM-YYYY date format (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3a — Only {date_format_correct_count}/{formula_count} formulas use DD-MMM-YYYY date format")

            if amount_ok:
                print(f"PASS: Component 3b — All {formula_count} formulas use #,##0.00 amount format (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3b — Only {amount_format_correct_count}/{formula_count} formulas use #,##0.00 amount format")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
