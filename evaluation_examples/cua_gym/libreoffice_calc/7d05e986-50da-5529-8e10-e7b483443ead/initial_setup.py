"""
Initial Setup: Transaction spreadsheet with empty column E (no narrative formulas yet)
Task ID: osworld_calc_text_format_number_008
Domain: libreoffice_calc

Creates a spreadsheet with Transaction ID (A), Date (B), Category (C), Amount (D).
Column E is intentionally EMPTY — the agent's task is to fill it with TEXT() formulas
like: ="On "&TEXT(B2,"DD-MMM-YYYY")&", a "&C2&" transaction of $"&TEXT(D2,"#,##0.00")&" was recorded."
"""

import os
import shlex
import subprocess
import time
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_text_format_number_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Transactions ---
    ws = wb.active
    ws.title = 'Transactions'

    # All 5 column headers (A-E)
    headers = ['Transaction ID', 'Date', 'Category', 'Amount', 'Narrative']
    header_font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic personal finance transaction data
    # Column E intentionally left empty — that is the task for the agent
    data = [
        ['TXN-0001', date(2025, 1, 7),  'Groceries',     124.35],
        ['TXN-0002', date(2025, 1, 12), 'Utilities',      89.50],
        ['TXN-0003', date(2025, 1, 15), 'Dining',         67.80],
        ['TXN-0004', date(2025, 1, 19), 'Transport',      45.20],
        ['TXN-0005', date(2025, 1, 25), 'Healthcare',    310.00],
        ['TXN-0006', date(2025, 2, 3),  'Entertainment',  55.99],
        ['TXN-0007', date(2025, 2, 8),  'Groceries',      98.45],
        ['TXN-0008', date(2025, 2, 14), 'Clothing',      230.00],
        ['TXN-0009', date(2025, 2, 20), 'Utilities',      74.15],
        ['TXN-0010', date(2025, 2, 27), 'Dining',         42.60],
        ['TXN-0011', date(2025, 3, 5),  'Transport',      33.75],
        ['TXN-0012', date(2025, 3, 11), 'Healthcare',    150.00],
        ['TXN-0013', date(2025, 3, 18), 'Entertainment',  88.25],
        ['TXN-0014', date(2025, 3, 22), 'Groceries',     112.90],
        ['TXN-0015', date(2025, 3, 29), 'Clothing',      199.99],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Transaction ID
        ws.cell(row=r, column=2, value=row_data[1])   # Date (date object -> Excel serial)
        ws.cell(row=r, column=2).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=3, value=row_data[2])   # Category
        ws.cell(row=r, column=4, value=row_data[3])   # Amount
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        # Column E (column 5) is intentionally left EMPTY

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 65

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
