"""
Initial Setup: temperatures.xlsx with daily temperature readings (some blank)
Task ID: osworld_multi_apps_calc_vscode_004
Domain: libreoffice_calc + vs-code (multi-app)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_calc_vscode_004'
OUTPUT = f'{WORKDIR}/temperatures.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    os.makedirs(WORKDIR, exist_ok=True)

    wb = openpyxl.Workbook()

    # --- Sheet: Temperatures ---
    ws = wb.active
    ws.title = "Temperatures"

    # Headers
    ws.cell(row=1, column=1, value="Date")
    ws.cell(row=1, column=2, value="Temperature")

    # Daily temperature data — some entries intentionally left blank
    # Blank rows: 3, 7, 12, 16 (Temperature column)
    data = [
        ("2025-01-01", 22.5),
        ("2025-01-02", 19.3),
        ("2025-01-03", None),   # blank
        ("2025-01-04", 24.7),
        ("2025-01-05", 21.0),
        ("2025-01-06", 18.6),
        ("2025-01-07", None),   # blank
        ("2025-01-08", 23.4),
        ("2025-01-09", 20.8),
        ("2025-01-10", 17.2),
        ("2025-01-11", 25.1),
        ("2025-01-12", None),   # blank
        ("2025-01-13", 22.0),
        ("2025-01-14", 19.7),
        ("2025-01-15", 21.5),
        ("2025-01-16", None),   # blank
        ("2025-01-17", 23.8),
        ("2025-01-18", 20.3),
        ("2025-01-19", 18.9),
        ("2025-01-20", 24.2),
    ]

    for r, (date_str, temp) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=date_str)
        if temp is not None:
            ws.cell(row=r, column=2, value=temp)
        # blank temperature cells are left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the xlsx in LibreOffice Calc first, then VSCode
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    # Also open VSCode pointed at the Desktop so the agent can create the script
    launch_gui('code "/home/user/Desktop"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc and VSCode with DISPLAY=:0')


create_initial()
