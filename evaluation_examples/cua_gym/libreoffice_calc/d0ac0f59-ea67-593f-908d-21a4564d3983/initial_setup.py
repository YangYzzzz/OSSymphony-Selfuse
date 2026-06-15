"""
Initial Setup: OFFSET function dynamic lookup task
Task ID: calc_fma_offset_007
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_offset_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: DynamicLookup ---
    ws = wb.active
    ws.title = 'DynamicLookup'

    # Offset parameters
    ws['D1'] = 2   # row offset from A5
    ws['E1'] = 3   # column offset from A5

    # Labels for the offset parameters
    ws['C1'] = 'Row Offset:'
    ws['C1'].font = Font(bold=True)
    ws['D1'].alignment = Alignment(horizontal='center')

    ws['F1'] = 'Col Offset:'
    ws['F1'].font = Font(bold=True)
    ws['E1'].alignment = Alignment(horizontal='center')

    # Label for the result cell
    ws['A3'] = 'Dynamic Result:'
    ws['A3'].font = Font(bold=True)
    # B2 is intentionally LEFT EMPTY — this is where the agent must enter the formula
    # B3 has label per task context
    ws['B3'] = 'Dynamic Result:'
    ws['B3'].font = Font(bold=True)

    # Sales data table header row (row 5)
    headers = ['Region', 'Jan', 'Feb', 'Mar', 'Apr']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center')

    # Sales data rows 6-9
    data = [
        ['North', 5000, 5200, 4800, 5500],
        ['South', 4200, 4000, 4500, 4300],
        ['East',  3800, 4100, 3900, 4200],
        ['West',  4900, 5100, 4700, 5300],
    ]
    for r, row_data in enumerate(data, 6):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal='center' if c > 1 else 'left')

    # Column widths for readability
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
