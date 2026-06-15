"""
Reward Script: Use OFFSET function in cell B2 for dynamic lookup
Task ID: calc_fma_offset_007
Domain: libreoffice_calc

Scoring Rubric:
- Component 1: Cell B2 contains an OFFSET formula (0.5 pts)
  - FAILS on initial (B2 is empty) -> PASSES on golden (B2 has formula)
- Component 2: OFFSET formula has correct structure: anchor=$A$5, row_offset=D1, col_offset=E1 (0.5 pts)
  - FAILS on initial (B2 is empty) -> PASSES on golden (formula is =OFFSET($A$5,D1,E1))

Total: 1.0
"""

import os
import openpyxl
import re

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_offset_007'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires inserting =OFFSET($A$5,D1,E1) in cell B2 of the DynamicLookup sheet.
    Initial state: B2 is empty.
    Golden state: B2 contains =OFFSET($A$5,D1,E1).
    """
    total_score = 0.0

    # Load the workbook (formula mode, not data_only, so we can check formula strings)
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: check DynamicLookup sheet exists
    if 'DynamicLookup' not in wb.sheetnames:
        print("CRITICAL: Sheet 'DynamicLookup' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['DynamicLookup']

    # Component 1: Cell B2 contains an OFFSET formula (0.5 points)
    # This FAILS on initial (B2 is empty) and PASSES on golden (B2 has OFFSET formula)
    try:
        b2_value = ws.cell(row=2, column=2).value
        if b2_value is not None and isinstance(b2_value, str) and 'OFFSET' in b2_value.upper():
            print(f"PASS: Component 1 — Cell B2 contains an OFFSET formula: {repr(b2_value)} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Expected OFFSET formula in B2, found: {repr(b2_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check B2: {e}")

    # Component 2: OFFSET formula has correct structure with $A$5 anchor, D1 row offset, E1 col offset (0.5 points)
    # This FAILS on initial (B2 is empty) and PASSES on golden (formula is =OFFSET($A$5,D1,E1))
    try:
        b2_value = ws.cell(row=2, column=2).value
        if b2_value is not None and isinstance(b2_value, str):
            # Normalize the formula for comparison: strip whitespace and convert to uppercase
            formula_normalized = b2_value.strip().upper().replace(' ', '')
            # Check for the correct OFFSET structure: =OFFSET($A$5,D1,E1)
            # Allow variations like OFFSET($A$5,D1,E1) with or without leading =
            # The anchor must be $A$5, row offset from D1, col offset from E1
            if re.search(r'OFFSET\(\$A\$5,D1,E1\)', formula_normalized):
                print(f"PASS: Component 2 — OFFSET formula has correct structure ($A$5 anchor, D1 row offset, E1 col offset): {repr(b2_value)} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — Expected OFFSET($A$5,D1,E1) structure, found: {repr(b2_value)}")
                print(f"       Normalized: {formula_normalized}")
        else:
            print(f"FAIL: Component 2 — B2 is empty or not a formula string, cannot verify OFFSET structure")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not verify OFFSET structure: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
