"""
Reward Script: Personal Monthly Budget Tracker
Task ID: calc_gen_personal_026
Domain: libreoffice_calc
Scoring:
  Component 1: Budget structure (title, income section, expense section, net/rate rows) — 0.30
  Component 2: Key formulas present (SUM income, SUM expenses, net savings, savings rate) — 0.25
  Component 3: Dollar number format on amount cells, percentage format on savings rate — 0.20
  Component 4: Conditional formatting on B23 (red fill when < 0) — 0.15
  Component 5: Pie chart present on Budget sheet — 0.10
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_personal_026'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Budget' sheet must exist
    if 'Budget' not in wb.sheetnames:
        print("FAIL: 'Budget' sheet not found — cannot verify task")
        print("\nScore: 0.0/1.0")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Budget']

    # -----------------------------------------------------------------------
    # Component 1: Budget structure (0.30 points)
    # Checks:
    #   - A1 contains 'Monthly Budget Tracker' (title)
    #   - A1 is bold (formatting of title)
    #   - Income categories present (Salary, Freelance, Other in rows 3-5)
    #   - 'Total Income' label present in A8
    #   - Expense categories present (Rent in A10, at least 10 rows A10:A19)
    #   - 'Total Expenses' label in A21
    #   - 'Net Savings' label in A23
    #   - 'Savings Rate' label in A24
    # -----------------------------------------------------------------------
    try:
        structure_checks = 0
        structure_total = 8

        # Title
        a1_val = ws['A1'].value
        if a1_val and 'Monthly Budget Tracker' in str(a1_val):
            structure_checks += 1
            print("PASS: A1 contains 'Monthly Budget Tracker'")
        else:
            print(f"FAIL: A1 expected 'Monthly Budget Tracker', found {repr(a1_val)}")

        # Title is bold
        if ws['A1'].font.bold:
            structure_checks += 1
            print("PASS: A1 title is bold")
        else:
            print(f"FAIL: A1 title is not bold (bold={ws['A1'].font.bold})")

        # Income categories
        income_cats = ['Salary', 'Freelance', 'Other']
        income_vals = [ws.cell(row=r, column=1).value for r in range(3, 8)]
        found_income = sum(1 for cat in income_cats if cat in income_vals)
        if found_income >= 2:
            structure_checks += 1
            print(f"PASS: Income categories found ({found_income}/{len(income_cats)}): {income_vals}")
        else:
            print(f"FAIL: Income categories missing. Expected {income_cats}, found {income_vals}")

        # Total Income label
        a8_val = ws['A8'].value
        if a8_val and 'Income' in str(a8_val):
            structure_checks += 1
            print(f"PASS: A8 contains income total label: {repr(a8_val)}")
        else:
            print(f"FAIL: A8 expected total income label, found {repr(a8_val)}")

        # Expense categories (at least 8 of 10 expected in rows 10-20)
        expense_cats = ['Rent', 'Food', 'Transport', 'Utilities', 'Entertainment', 'Health', 'Insurance', 'Savings', 'Misc', 'Other']
        expense_vals = [ws.cell(row=r, column=1).value for r in range(10, 21)]
        found_expenses = sum(1 for cat in expense_cats if cat in expense_vals)
        if found_expenses >= 8:
            structure_checks += 1
            print(f"PASS: Expense categories found ({found_expenses}/{len(expense_cats)})")
        else:
            print(f"FAIL: Expense categories missing. Found only {found_expenses}/{len(expense_cats)}: {expense_vals}")

        # Total Expenses label
        a21_val = ws['A21'].value
        if a21_val and 'Expense' in str(a21_val):
            structure_checks += 1
            print(f"PASS: A21 contains expense total label: {repr(a21_val)}")
        else:
            print(f"FAIL: A21 expected total expense label, found {repr(a21_val)}")

        # Net Savings label
        a23_val = ws['A23'].value
        if a23_val and 'Saving' in str(a23_val):
            structure_checks += 1
            print(f"PASS: A23 contains net savings label: {repr(a23_val)}")
        else:
            print(f"FAIL: A23 expected 'Net Savings' label, found {repr(a23_val)}")

        # Savings Rate label
        a24_val = ws['A24'].value
        if a24_val and 'Rate' in str(a24_val):
            structure_checks += 1
            print(f"PASS: A24 contains savings rate label: {repr(a24_val)}")
        else:
            print(f"FAIL: A24 expected 'Savings Rate' label, found {repr(a24_val)}")

        component1_score = 0.30 * (structure_checks / structure_total)
        print(f"Component 1: {structure_checks}/{structure_total} structure checks passed — {component1_score:.2f} pts")
        total_score += component1_score

    except Exception as e:
        print(f"ERROR: Component 1 (structure) — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Key formulas (0.25 points)
    # B8 = =SUM(B3:B7)
    # B21 = =SUM(B10:B20)
    # B23 = =B8-B21
    # B24 = =B23/B8
    # -----------------------------------------------------------------------
    try:
        formula_checks = 0
        formula_total = 4

        # B8: SUM formula for income
        b8_val = ws['B8'].value
        if b8_val and isinstance(b8_val, str) and 'SUM' in b8_val.upper() and 'B' in b8_val.upper():
            formula_checks += 1
            print(f"PASS: B8 has SUM formula: {repr(b8_val)}")
        else:
            print(f"FAIL: B8 expected SUM formula, found {repr(b8_val)}")

        # B21: SUM formula for expenses
        b21_val = ws['B21'].value
        if b21_val and isinstance(b21_val, str) and 'SUM' in b21_val.upper() and 'B' in b21_val.upper():
            formula_checks += 1
            print(f"PASS: B21 has SUM formula: {repr(b21_val)}")
        else:
            print(f"FAIL: B21 expected SUM formula, found {repr(b21_val)}")

        # B23: Net savings = income - expenses
        b23_val = ws['B23'].value
        if b23_val and isinstance(b23_val, str) and 'B8' in b23_val and 'B21' in b23_val and '-' in b23_val:
            formula_checks += 1
            print(f"PASS: B23 has net savings formula: {repr(b23_val)}")
        else:
            print(f"FAIL: B23 expected =B8-B21 formula, found {repr(b23_val)}")

        # B24: Savings rate = savings / income
        b24_val = ws['B24'].value
        if b24_val and isinstance(b24_val, str) and 'B23' in b24_val and 'B8' in b24_val and '/' in b24_val:
            formula_checks += 1
            print(f"PASS: B24 has savings rate formula: {repr(b24_val)}")
        else:
            print(f"FAIL: B24 expected =B23/B8 formula, found {repr(b24_val)}")

        component2_score = 0.25 * (formula_checks / formula_total)
        print(f"Component 2: {formula_checks}/{formula_total} formula checks passed — {component2_score:.2f} pts")
        total_score += component2_score

    except Exception as e:
        print(f"ERROR: Component 2 (formulas) — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Number formats (0.20 points)
    # Amount cells in B column: $#,##0.00
    # Savings rate B24: 0.00% (percentage format)
    # -----------------------------------------------------------------------
    try:
        format_checks = 0
        format_total = 2

        # Check dollar format on at least 5 of the income/expense amount cells
        dollar_cells = [3, 4, 5, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 23]
        dollar_count = 0
        for row in dollar_cells:
            c = ws.cell(row=row, column=2)
            fmt = c.number_format
            if fmt and ('$' in fmt or '0.00' in fmt) and '#' in fmt:
                dollar_count += 1

        if dollar_count >= 8:
            format_checks += 1
            print(f"PASS: Dollar format ($#,##0.00) applied to {dollar_count}/{len(dollar_cells)} amount cells")
        else:
            print(f"FAIL: Dollar format only on {dollar_count}/{len(dollar_cells)} amount cells (need >= 8)")

        # Check percentage format on B24
        b24_fmt = ws['B24'].number_format
        if b24_fmt and ('%' in b24_fmt):
            format_checks += 1
            print(f"PASS: B24 has percentage format: {repr(b24_fmt)}")
        else:
            print(f"FAIL: B24 expected percentage format, found {repr(b24_fmt)}")

        component3_score = 0.20 * (format_checks / format_total)
        print(f"Component 3: {format_checks}/{format_total} format checks passed — {component3_score:.2f} pts")
        total_score += component3_score

    except Exception as e:
        print(f"ERROR: Component 3 (number formats) — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional formatting on B23 — red fill when < 0 (0.15 points)
    # -----------------------------------------------------------------------
    try:
        cf_rules = ws.conditional_formatting
        cf_found = False
        cf_red_fill = False
        cf_correct_range = False

        for cf in cf_rules:
            cf_str = str(cf)
            # Check if the CF applies to B23 or a range including B23
            if 'B23' in cf_str:
                cf_correct_range = True
                for rule in cf.rules:
                    # Check for cellIs lessThan 0
                    rule_type = getattr(rule, 'type', None)
                    rule_op = getattr(rule, 'operator', None)
                    rule_formula = getattr(rule, 'formula', [])
                    if rule_type == 'cellIs' and rule_op == 'lessThan' and rule_formula and '0' in str(rule_formula):
                        cf_found = True
                        # Check for red fill
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                fill_color = rule.dxf.fill.fgColor.rgb
                                if fill_color and ('FF0000' in fill_color.upper() or fill_color.upper() == 'FFFF0000'):
                                    cf_red_fill = True
                            except Exception:
                                pass

        if cf_correct_range and cf_found and cf_red_fill:
            print(f"PASS: Conditional formatting on B23 — red fill when value < 0")
            total_score += 0.15
        elif cf_correct_range and cf_found:
            print(f"PASS (partial): Conditional formatting on B23 found with lessThan rule, but fill color not verified as red")
            total_score += 0.10
        elif cf_correct_range:
            print(f"FAIL: Conditional formatting applied to B23 but rule type/operator not matching (need cellIs lessThan 0)")
        else:
            print(f"FAIL: No conditional formatting found on B23")

    except Exception as e:
        print(f"ERROR: Component 4 (conditional formatting) — {e}")

    # -----------------------------------------------------------------------
    # Component 5: Pie chart on Budget sheet (0.10 points)
    # -----------------------------------------------------------------------
    try:
        charts = ws._charts
        pie_charts = [c for c in charts if type(c).__name__ == 'PieChart']

        if len(pie_charts) >= 1:
            print(f"PASS: Pie chart found on Budget sheet ({len(pie_charts)} pie chart(s))")
            # Check if it has data (at least 1 series referencing expense data)
            pie = pie_charts[0]
            if len(pie.series) >= 1:
                total_score += 0.10
                print(f"PASS: Pie chart has {len(pie.series)} data series")
            else:
                print(f"FAIL: Pie chart has no data series")
                total_score += 0.05
        else:
            if len(charts) >= 1:
                print(f"FAIL: Chart found but not a pie chart (found: {[type(c).__name__ for c in charts]})")
            else:
                print(f"FAIL: No charts found on Budget sheet")

    except Exception as e:
        print(f"ERROR: Component 5 (pie chart) — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score:.4f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
