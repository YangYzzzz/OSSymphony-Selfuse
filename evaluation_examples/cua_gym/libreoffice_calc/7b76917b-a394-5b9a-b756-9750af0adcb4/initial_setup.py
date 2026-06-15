"""
Initial Setup: Sensor readings spreadsheet with blank gaps (no chart)
Task ID: calc_chart_blank_gaps_035
Domain: libreoffice_calc
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_blank_gaps_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Readings ---
    ws = wb.active
    ws.title = 'Readings'

    # Headers
    ws['A1'] = 'Time'
    ws['B1'] = 'Sensor Value'

    # Data rows — rows 4 and 5 are blank (sensor offline at 10:00 and 11:00)
    data = [
        ('08:00', 42.5),
        ('09:00', 44.1),
        ('10:00', None),   # blank — sensor offline
        ('11:00', None),   # blank — sensor offline
        ('12:00', 47.8),
        ('13:00', 49.2),
        ('14:00', 51.0),
        ('15:00', 48.6),
    ]

    for r, (time_val, sensor_val) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=time_val)
        ws.cell(row=r, column=2, value=sensor_val)

    # Column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16

    # NO charts in the initial file (task asks agent to create one)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets:', wb.sheetnames)
    print('Rows of data: 8 (including 2 blank sensor values at 10:00 and 11:00)')


create_initial()
