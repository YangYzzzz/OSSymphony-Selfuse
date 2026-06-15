"""
Reward Script: Create a line chart from sensor data configured to interpolate blank cells
Task ID: calc_chart_blank_gaps_035
Domain: libreoffice_calc
Scoring:
  Component 1: A LineChart exists on 'Readings' sheet (0.30 pts)
  Component 2: Chart title is 'Sensor Readings Over Time' (0.30 pts)
  Component 3: Series data range covers rows 2-9 of columns A and B (0.20 pts)
  Component 4: dispBlanksAs is set to 'span' (interpolate over blanks) (0.20 pts)
Total: 1.0
"""

import os
import zipfile
import re

import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_chart_blank_gaps_035'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must be loadable
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Readings' sheet must exist
    if 'Readings' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Readings' not found. Sheets present:", wb.sheetnames)
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Readings']

    # Component 1: A LineChart exists on the 'Readings' sheet (0.30 points)
    # On the initial file there are 0 charts; on the golden file there is 1 LineChart.
    try:
        from openpyxl.chart import LineChart as OxlLineChart
        charts = ws._charts
        line_charts = [c for c in charts if isinstance(c, OxlLineChart)]
        if len(line_charts) >= 1:
            print(f"PASS: Component 1 — LineChart found on 'Readings' sheet ({len(line_charts)} chart(s)) (0.30 pts)")
            total_score += 0.30
            chart = line_charts[0]
        else:
            print(f"FAIL: Component 1 — No LineChart found on 'Readings' sheet (charts found: {[type(c).__name__ for c in charts]})")
            # No chart means we cannot evaluate components 2-4 against chart properties
            chart = None
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        chart = None

    # Component 2: Chart title is 'Sensor Readings Over Time' (0.30 points)
    # This requires the chart to exist (component 1 must have passed).
    if chart is not None:
        try:
            # Navigate the openpyxl title object to extract the text
            title_text = None
            try:
                title_text = chart.title.tx.rich.p[0].r[0].t
            except Exception:
                pass
            # Fallback: check string representation
            if title_text is None:
                try:
                    title_str = str(chart.title)
                    match = re.search(r"t='([^']*)'", title_str)
                    if match:
                        title_text = match.group(1)
                except Exception:
                    pass

            expected_title = "Sensor Readings Over Time"
            if title_text and title_text.strip() == expected_title:
                print(f"PASS: Component 2 — Chart title is '{title_text}' (0.30 pts)")
                total_score += 0.30
            else:
                print(f"FAIL: Component 2 — Expected title '{expected_title}', found: '{title_text}'")
        except Exception as e:
            print(f"ERROR: Component 2 — {e}")

    # Component 3: Series data covers the expected range A1:B9 (0.20 points)
    # Data range: values from 'Readings'!$B$2:$B$9, categories from 'Readings'!$A$2:$A$9
    if chart is not None:
        try:
            found_val_range = False
            found_cat_range = False
            if len(chart.series) >= 1:
                s = chart.series[0]
                # Check value reference
                if s.val and s.val.numRef:
                    val_ref = s.val.numRef.f
                    # Accept any reference that covers rows 2-9 of column B
                    # Canonical form: 'Readings'!$B$2:$B$9
                    if 'B' in val_ref and '2' in val_ref and '9' in val_ref:
                        found_val_range = True
                # Check category reference
                if s.cat:
                    cat_ref_str = str(s.cat)
                    # Try to extract formula string
                    cat_formula = None
                    try:
                        if s.cat.numRef:
                            cat_formula = s.cat.numRef.f
                        elif s.cat.strRef:
                            cat_formula = s.cat.strRef.f
                    except Exception:
                        cat_formula = cat_ref_str
                    if cat_formula and 'A' in str(cat_formula) and '2' in str(cat_formula) and '9' in str(cat_formula):
                        found_cat_range = True

            if found_val_range and found_cat_range:
                print(f"PASS: Component 3 — Series data range covers rows 2-9 columns A-B (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 3 — Data range check failed. val_range_ok={found_val_range}, cat_range_ok={found_cat_range}")
        except Exception as e:
            print(f"ERROR: Component 3 — {e}")

    # Component 4: dispBlanksAs = 'span' (interpolate over blank cells) (0.20 points)
    # The XML attribute <dispBlanksAs val="span"/> must be present in the chart XML.
    # 'span' = bridge/interpolate over blanks; 'gap' = show breaks; 'zero' = treat as zero.
    # This is the key requirement: the chart must NOT show gaps for missing 10:00 and 11:00 values.
    try:
        with zipfile.ZipFile(file_path, 'r') as z:
            chart_xml_files = [n for n in z.namelist() if re.match(r'xl/charts/chart\d+\.xml', n)]
            if not chart_xml_files:
                print("FAIL: Component 4 — No chart XML file found in the workbook")
            else:
                # Check each chart XML for dispBlanksAs=span
                found_span = False
                for chart_file in chart_xml_files:
                    with z.open(chart_file) as f:
                        xml_content = f.read().decode('utf-8')
                    if 'dispBlanksAs' in xml_content:
                        # Extract the value
                        match = re.search(r'dispBlanksAs\s+val=["\']([^"\']+)["\']', xml_content)
                        if match:
                            val = match.group(1)
                            if val == 'span':
                                found_span = True
                                print(f"PASS: Component 4 — dispBlanksAs='span' found (interpolate over blanks) (0.20 pts)")
                            else:
                                print(f"FAIL: Component 4 — dispBlanksAs='{val}' (expected 'span' for interpolation)")
                        else:
                            print("FAIL: Component 4 — dispBlanksAs attribute found but value not extractable")
                    else:
                        print("FAIL: Component 4 — dispBlanksAs not present in chart XML (blanks will show as gaps)")

                if found_span:
                    total_score += 0.20
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
