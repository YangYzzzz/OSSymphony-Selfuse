"""
Initial Setup: Create a spreadsheet with inventory items where C2:C20 is empty,
ready for the agent to add custom validation (multiples of 5).
Task ID: calc_nrv_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    # --- Headers ---
    headers = ['Item', 'Description', 'Quantity (multiples of 5)']
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Item data (A2:B20) with realistic inventory items, C column empty ---
    items = [
        ('WH-1001', 'Steel Bolts M8x40mm'),
        ('WH-1002', 'Copper Wire Spool 2.5mm'),
        ('WH-1003', 'PVC Pipe Connector T-Joint'),
        ('WH-1004', 'LED Panel Light 60W'),
        ('WH-1005', 'Rubber Gasket Set 50mm'),
        ('WH-1006', 'Aluminum Sheet 1.5mm 4x8ft'),
        ('WH-1007', 'Stainless Steel Hinge 3-inch'),
        ('WH-1008', 'Concrete Anchor Bolt 10mm'),
        ('WH-1009', 'Fiberglass Insulation Roll R-19'),
        ('WH-1010', 'Brass Compression Fitting 15mm'),
        ('WH-1011', 'Galvanized Nail Box 3-inch'),
        ('WH-1012', 'Silicone Sealant Tube 300ml'),
        ('WH-1013', 'Plywood Board 18mm 4x8ft'),
        ('WH-1014', 'Electrical Conduit EMT 3/4-inch'),
        ('WH-1015', 'Carbon Steel Drill Bit Set'),
        ('WH-1016', 'Polyethylene Tarp 10x12ft'),
        ('WH-1017', 'Ceramic Floor Tile 12x12 Beige'),
        ('WH-1018', 'Hydraulic Hose 1/2-inch 50ft'),
        ('WH-1019', 'Safety Harness Full-Body D-Ring'),
    ]

    for r, (item_code, desc) in enumerate(items, 2):
        ws.cell(row=r, column=1, value=item_code).border = thin_border
        ws.cell(row=r, column=2, value=desc).border = thin_border
        # Column C is intentionally empty - agent must add validation
        ws.cell(row=r, column=3).border = thin_border

    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 38
    ws.column_dimensions['C'].width = 28

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
