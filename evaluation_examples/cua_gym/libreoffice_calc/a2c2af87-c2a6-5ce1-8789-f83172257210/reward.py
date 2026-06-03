"""
Reward Script: Track fleet utilization - formulas, chart, and conditional formatting
Task ID: calc_ops_fleet_utilization_032
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.30): E2:E9 contain idle days formulas (=Bx-Cx-Dx)
  - Component 2 (0.30): F2:F9 contain utilization rate formulas (=Cx/Bx) formatted as percentage
  - Component 3 (0.20): G2:G9 contain availability rate formulas (=(Bx-Dx)/Bx) formatted as percentage
  - Component 4 (0.10): Stacked bar chart exists with title 'Fleet Utilization This Month'
  - Component 5 (0.10): Conditional formatting on F2:F9 with green/amber/red rules
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_fleet_utilization_032'


def normalize_formula(formula):
    """Normalize a formula string for comparison: uppercase, no spaces."""
    if not isinstance(formula, str):
        return ''
    return formula.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the FleetUtilization sheet exists
    if 'FleetUtilization' not in wb.sheetnames:
        print("CRITICAL: 'FleetUtilization' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['FleetUtilization']

    # Component 1: E2:E9 contain idle days formulas =Bx-Cx-Dx (0.30 points)
    # These cells are empty in the initial file; formulas must be added by the agent
    try:
        idle_correct = 0
        idle_errors = []
        for row in range(2, 10):
            cell_e = ws.cell(row=row, column=5)  # Column E
            val = cell_e.value
            if val is None:
                idle_errors.append(f"E{row}: None (empty)")
                continue
            val_norm = normalize_formula(val)
            # Expected: =Bx-Cx-Dx (e.g., =B2-C2-D2)
            expected = f"=B{row}-C{row}-D{row}"
            expected_norm = normalize_formula(expected)
            if val_norm == expected_norm:
                idle_correct += 1
            else:
                idle_errors.append(f"E{row}: expected {expected}, found {repr(val)}")

        if idle_correct == 8:
            print(f"PASS: Component 1 — All 8 idle days formulas correct in E2:E9 (0.30 pts)")
            total_score += 0.30
        elif idle_correct >= 4:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {idle_correct}/8 idle days formulas correct (0.15 pts)")
            print(f"  Issues: {idle_errors}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {idle_correct}/8 idle days formulas correct")
            print(f"  Issues: {idle_errors}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: F2:F9 contain utilization rate formulas =Cx/Bx AND formatted as percentage (0.30 points)
    # These cells are empty in the initial file
    try:
        util_formula_correct = 0
        util_format_correct = 0
        util_errors = []
        for row in range(2, 10):
            cell_f = ws.cell(row=row, column=6)  # Column F
            val = cell_f.value
            if val is None:
                util_errors.append(f"F{row}: None (empty)")
                continue
            val_norm = normalize_formula(val)
            expected = f"=C{row}/B{row}"
            expected_norm = normalize_formula(expected)
            if val_norm == expected_norm:
                util_formula_correct += 1
            else:
                util_errors.append(f"F{row}: expected {expected}, found {repr(val)}")
            # Check number format is percentage
            fmt = cell_f.number_format
            if fmt and '%' in fmt:
                util_format_correct += 1
            else:
                util_errors.append(f"F{row}: number_format={repr(fmt)} (expected % format)")

        if util_formula_correct == 8 and util_format_correct == 8:
            print(f"PASS: Component 2 — All 8 utilization rate formulas correct + percentage format in F2:F9 (0.30 pts)")
            total_score += 0.30
        elif util_formula_correct == 8:
            partial = 0.20
            print(f"PARTIAL: Component 2 — Formulas correct but only {util_format_correct}/8 have percentage format (0.20 pts)")
            print(f"  Issues: {util_errors}")
            total_score += partial
        elif util_formula_correct >= 4:
            partial = 0.10
            print(f"PARTIAL: Component 2 — {util_formula_correct}/8 formulas correct (0.10 pts)")
            print(f"  Issues: {util_errors}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {util_formula_correct}/8 utilization rate formulas correct")
            print(f"  Issues: {util_errors}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: G2:G9 contain availability rate formulas =(Bx-Dx)/Bx AND formatted as percentage (0.20 points)
    # These cells are empty in the initial file
    try:
        avail_formula_correct = 0
        avail_format_correct = 0
        avail_errors = []
        for row in range(2, 10):
            cell_g = ws.cell(row=row, column=7)  # Column G
            val = cell_g.value
            if val is None:
                avail_errors.append(f"G{row}: None (empty)")
                continue
            val_norm = normalize_formula(val)
            expected = f"=(B{row}-D{row})/B{row}"
            expected_norm = normalize_formula(expected)
            if val_norm == expected_norm:
                avail_formula_correct += 1
            else:
                avail_errors.append(f"G{row}: expected {expected}, found {repr(val)}")
            # Check number format is percentage
            fmt = cell_g.number_format
            if fmt and '%' in fmt:
                avail_format_correct += 1
            else:
                avail_errors.append(f"G{row}: number_format={repr(fmt)} (expected % format)")

        if avail_formula_correct == 8 and avail_format_correct == 8:
            print(f"PASS: Component 3 — All 8 availability rate formulas correct + percentage format in G2:G9 (0.20 pts)")
            total_score += 0.20
        elif avail_formula_correct == 8:
            partial = 0.12
            print(f"PARTIAL: Component 3 — Formulas correct but only {avail_format_correct}/8 have percentage format (0.12 pts)")
            print(f"  Issues: {avail_errors}")
            total_score += partial
        elif avail_formula_correct >= 4:
            partial = 0.07
            print(f"PARTIAL: Component 3 — {avail_formula_correct}/8 availability formulas correct (0.07 pts)")
            print(f"  Issues: {avail_errors}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {avail_formula_correct}/8 availability formulas correct")
            print(f"  Issues: {avail_errors}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Stacked bar chart exists with title 'Fleet Utilization This Month' (0.10 points)
    # No chart exists in the initial file
    try:
        charts = ws._charts
        if len(charts) == 0:
            print("FAIL: Component 4 — No chart found in FleetUtilization sheet")
        else:
            chart = charts[0]
            # Check it's a BarChart with stacked grouping
            is_bar = type(chart).__name__ == 'BarChart'
            is_stacked = getattr(chart, 'grouping', '') == 'stacked'
            # Check title contains expected text
            title_ok = False
            try:
                if chart.title is not None:
                    # Navigate the title object to get text
                    title_text = ''
                    t = chart.title
                    if hasattr(t, 'tx') and t.tx:
                        tx = t.tx
                        if hasattr(tx, 'rich') and tx.rich:
                            for para in tx.rich.p:
                                for run in para.r:
                                    title_text += run.t
                    title_ok = 'Fleet Utilization This Month' in title_text
                    print(f"  Chart title text found: {repr(title_text)}")
            except Exception as te:
                print(f"  Title check error: {te}")

            if is_bar and is_stacked and title_ok:
                print(f"PASS: Component 4 — Stacked bar chart with correct title found (0.10 pts)")
                total_score += 0.10
            elif is_bar and is_stacked:
                partial = 0.05
                print(f"PARTIAL: Component 4 — Stacked bar chart found but title incorrect (0.05 pts)")
                total_score += partial
            elif len(charts) >= 1:
                partial = 0.03
                print(f"PARTIAL: Component 4 — Chart found but not stacked bar type (0.03 pts)")
                print(f"  Chart type: {type(chart).__name__}, grouping: {getattr(chart, 'grouping', 'N/A')}")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — Chart found but requirements not met")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on F2:F9 with at least 2 rules (green/amber/red) (0.10 points)
    # No conditional formatting exists in the initial file
    try:
        cf_found = False
        cf_rule_count = 0
        cf_range_found = False

        for cf_range in ws.conditional_formatting:
            range_str = str(cf_range)
            # Check if conditional formatting targets F2:F9
            if 'F2' in range_str and 'F9' in range_str:
                cf_range_found = True
                rules = ws.conditional_formatting[cf_range]
                cf_rule_count = len(list(rules))

        if cf_range_found and cf_rule_count >= 2:
            print(f"PASS: Component 5 — Conditional formatting on F2:F9 with {cf_rule_count} rules (0.10 pts)")
            total_score += 0.10
        elif cf_range_found:
            partial = 0.05
            print(f"PARTIAL: Component 5 — CF on F2:F9 found but only {cf_rule_count} rule(s) (0.05 pts)")
            total_score += partial
        else:
            # Check if any CF exists at all on F column
            any_cf_f = False
            for cf_range in ws.conditional_formatting:
                if 'F' in str(cf_range):
                    any_cf_f = True
                    break
            if any_cf_f:
                partial = 0.03
                print(f"PARTIAL: Component 5 — CF found on F column but not F2:F9 range (0.03 pts)")
                total_score += partial
            else:
                print("FAIL: Component 5 — No conditional formatting found on F2:F9")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 4)}/1.0")
    print(f"REWARD: {round(final_score, 4)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
