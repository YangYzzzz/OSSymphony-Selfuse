"""
Initial Setup: Fleet Utilization Tracker
Task ID: calc_ops_fleet_utilization_032
Domain: libreoffice_calc
Creates a FleetUtilization spreadsheet with vehicle data (A-D filled, E-G empty).
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ops_fleet_utilization_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: FleetUtilization ---
    ws = wb.active
    ws.title = 'FleetUtilization'

    # Headers in row 1
    headers = [
        'Vehicle ID',          # A
        'Available Days',      # B
        'Days In Use',         # C
        'Days In Maintenance', # D
        'Days Idle',           # E — empty (formula to be added)
        'Utilization Rate',    # F — empty (formula to be added)
        'Availability Rate',   # G — empty (formula to be added)
    ]
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFFFF')
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # 8 vehicles — realistic fleet IDs, B=30 for all, varied C and D
    # Days In Use + Days In Maintenance <= 30
    vehicle_data = [
        # Vehicle ID,  Available, In Use, In Maintenance
        ('VH-001 Iveco Daily',      30, 24, 2),
        ('VH-002 Ford Transit',     30, 26, 1),
        ('VH-003 Mercedes Sprinter',30, 18, 6),
        ('VH-004 Renault Master',   30, 22, 3),
        ('VH-005 VW Crafter',       30, 15, 8),
        ('VH-006 Peugeot Boxer',    30, 27, 0),
        ('VH-007 Fiat Ducato',      30, 20, 5),
        ('VH-008 Citroen Relay',    30, 12, 4),
    ]

    data_align = Alignment(horizontal='center', vertical='center')
    data_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, (vid, avail, in_use, in_maint) in enumerate(vehicle_data, 2):
        # Column A: Vehicle ID
        cell_a = ws.cell(row=r, column=1, value=vid)
        cell_a.font = Font(name='Calibri', size=11)
        cell_a.alignment = Alignment(horizontal='left', vertical='center')
        cell_a.border = data_border

        # Column B: Available Days (30 for all)
        cell_b = ws.cell(row=r, column=2, value=avail)
        cell_b.font = Font(name='Calibri', size=11)
        cell_b.alignment = data_align
        cell_b.border = data_border

        # Column C: Days In Use
        cell_c = ws.cell(row=r, column=3, value=in_use)
        cell_c.font = Font(name='Calibri', size=11)
        cell_c.alignment = data_align
        cell_c.border = data_border

        # Column D: Days In Maintenance
        cell_d = ws.cell(row=r, column=4, value=in_maint)
        cell_d.font = Font(name='Calibri', size=11)
        cell_d.alignment = data_align
        cell_d.border = data_border

        # Columns E, F, G: intentionally empty (to be filled by agent)
        for col in [5, 6, 7]:
            cell = ws.cell(row=r, column=col, value=None)
            cell.border = data_border

    # Column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 17

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Row height for header
    ws.row_dimensions[1].height = 30

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheet: FleetUtilization')
    print('Rows: 8 vehicles (rows 2-9)')
    print('Columns A-D filled; E-G empty (awaiting formulas)')


create_initial()
