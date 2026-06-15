"""
Initial Setup: ABC Analysis on Product Catalog
Task ID: calc_sales_product_abc_016
Domain: libreoffice_calc

Creates a spreadsheet with 50 products sorted by Annual Revenue descending.
Columns D (Revenue %), E (Cumulative %), F (ABC Class) are intentionally left empty.
Total revenue sum in C52.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_product_abc_016'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'

    # --- Headers ---
    headers = ['SKU', 'Product Name', 'Annual Revenue', 'Revenue %', 'Cumulative %', 'ABC Class']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF', name='Calibri', size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Product data: 50 products sorted by Annual Revenue descending ---
    # Revenue range: $890,000 down to $2,400
    products = [
        ('SKU-001', 'UltraBook Pro 15',          890000),
        ('SKU-002', 'Gaming Laptop Elite',        742000),
        ('SKU-003', 'SmartTV 65" OLED',           698500),
        ('SKU-004', 'Enterprise Server X900',     634000),
        ('SKU-005', 'Wireless Headphones Pro',    521000),
        ('SKU-006', 'Digital Camera MX700',       487500),
        ('SKU-007', 'Tablet Air 12',              432000),
        ('SKU-008', 'Gaming Console Next',        398700),
        ('SKU-009', 'Smart Watch Series 5',       365000),
        ('SKU-010', 'Monitor 4K UHD 32"',         342000),
        ('SKU-011', 'Mechanical Keyboard TKL',    318000),
        ('SKU-012', 'Drone Explorer X1',          295000),
        ('SKU-013', 'VR Headset Plus',            271000),
        ('SKU-014', 'Portable SSD 2TB',           248500),
        ('SKU-015', 'Wireless Mouse Precision',   226000),
        ('SKU-016', 'Bluetooth Speaker Max',      208000),
        ('SKU-017', 'Network Switch 24-Port',     191000),
        ('SKU-018', 'Graphics Card RTX',          175500),
        ('SKU-019', 'NAS Storage 8-Bay',          162000),
        ('SKU-020', 'USB-C Hub 10-Port',          149000),
        ('SKU-021', 'Laser Printer Color',        137500),
        ('SKU-022', 'Webcam 4K Ultra',            125000),
        ('SKU-023', 'Desktop Mini PC',            113000),
        ('SKU-024', 'Smart Home Hub',             102000),
        ('SKU-025', 'Action Camera H10',           92400),
        ('SKU-026', 'Ergonomic Chair Pro',         83500),
        ('SKU-027', 'Standing Desk Electric',      75800),
        ('SKU-028', 'Noise Cancelling Earbuds',    68200),
        ('SKU-029', 'Router WiFi 6E',              61500),
        ('SKU-030', 'LED Desk Lamp Smart',         55000),
        ('SKU-031', 'Portable Projector Mini',     49200),
        ('SKU-032', 'Phone Charging Pad Wireless', 44100),
        ('SKU-033', 'External DVD Drive',          39600),
        ('SKU-034', 'Laptop Cooling Pad Pro',      35400),
        ('SKU-035', 'Screen Protector Glass Set',  31700),
        ('SKU-036', 'Cable Management Kit',        28300),
        ('SKU-037', 'HDMI Switch 4-Port',          25100),
        ('SKU-038', 'Microphone USB Condenser',    22400),
        ('SKU-039', 'Power Strip Smart 6-Port',    19800),
        ('SKU-040', 'Webcam Cover Privacy',        17500),
        ('SKU-041', 'Wrist Rest Ergonomic',        15400),
        ('SKU-042', 'Surge Protector 8-Outlet',    13600),
        ('SKU-043', 'Screen Cleaning Kit',         11900),
        ('SKU-044', 'USB Flash Drive 128GB',       10400),
        ('SKU-045', 'Cable Organizer Set',          8800),
        ('SKU-046', 'Monitor Stand Adjustable',     7200),
        ('SKU-047', 'Mouse Pad XL Gaming',          5700),
        ('SKU-048', 'Cable Clip Pack 50pcs',        4300),
        ('SKU-049', 'Screen Wipe Refills',          3200),
        ('SKU-050', 'Keyboard Dust Cover',          2400),
    ]

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 12

    # Write product data (D, E, F intentionally left empty)
    thin = Side(style='thin', color='FFCCCCCC')
    for r, (sku, name, revenue) in enumerate(products, 2):
        ws.cell(row=r, column=1, value=sku)
        ws.cell(row=r, column=2, value=name)
        cell_rev = ws.cell(row=r, column=3, value=revenue)
        cell_rev.number_format = '$#,##0'
        # Columns D (4), E (5), F (6) are empty — task requires filling these
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )
        # Alternate row shading
        if r % 2 == 0:
            row_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
            for c in range(1, 7):
                ws.cell(row=r, column=c).fill = row_fill

    # --- Row 52: Total Revenue label and SUM formula ---
    label_cell = ws.cell(row=52, column=2, value='Total Revenue')
    label_cell.font = Font(bold=True, name='Calibri', size=11)
    label_cell.alignment = Alignment(horizontal='right')

    total_cell = ws.cell(row=52, column=3, value='=SUM(C2:C51)')
    total_cell.number_format = '$#,##0'
    total_cell.font = Font(bold=True, name='Calibri', size=11)

    # Freeze pane at row 2 (keep headers visible)
    ws.freeze_panes = 'A2'

    # Row 1 height
    ws.row_dimensions[1].height = 22

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Products')
    print(f'  Products: {len(products)} rows (rows 2-51), sorted by Annual Revenue descending')
    print(f'  Revenue range: $2,400 to $890,000')
    print(f'  Columns D, E, F: intentionally empty (task will fill these)')
    print(f'  Total revenue: cell C52 = =SUM(C2:C51)')


create_initial()
