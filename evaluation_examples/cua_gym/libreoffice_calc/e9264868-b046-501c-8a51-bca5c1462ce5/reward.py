"""
Reward Script: ABC Analysis on Product Catalog
Task ID: calc_sales_product_abc_016
Domain: libreoffice_calc
Scoring:
  Component 1: Revenue % formulas in D2:D51 (=C{row}/$C$52)          — 0.25 pts
  Component 2: Cumulative % formulas in E2:E51 (=D2; =E{n-1}+D{n})   — 0.25 pts
  Component 3: ABC class IFS formulas in F2:F51                        — 0.25 pts
  Component 4: Pie chart present with SUMIF data source in H/I cols    — 0.25 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_product_abc_016'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet 'Products' exists
    if 'Products' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Products' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Products']

    # Component 1: Revenue % formulas in D2:D51 (0.25 points)
    # Each row should have =C{row}/$C$52 in column D
    try:
        d_correct = 0
        d_total = 50  # rows 2-51
        for row in range(2, 52):
            val = ws.cell(row=row, column=4).value
            if val is not None and isinstance(val, str):
                # Normalize formula: remove spaces, uppercase
                norm = val.upper().replace(' ', '')
                expected = f'=C{row}/$C$52'
                expected_norm = expected.upper().replace(' ', '')
                if norm == expected_norm:
                    d_correct += 1
        ratio = d_correct / d_total
        if ratio == 1.0:
            print(f"PASS: Component 1 — Revenue % formulas: all {d_correct}/{d_total} rows correct (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.8:
            partial = round(0.25 * ratio, 4)
            print(f"PARTIAL: Component 1 — Revenue % formulas: {d_correct}/{d_total} rows correct (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Revenue % formulas: only {d_correct}/{d_total} rows correct (expected =C{{row}}/$C$52)")
    except Exception as e:
        print(f"ERROR: Component 1 — Revenue % formulas: {e}")

    # Component 2: Cumulative % formulas in E2:E51 (0.25 points)
    # E2 = =D2; E3:E51 = =E{row-1}+D{row}
    try:
        e_correct = 0
        e_total = 50  # rows 2-51

        # Check E2 = =D2
        e2_val = ws.cell(row=2, column=5).value
        if e2_val is not None and isinstance(e2_val, str):
            norm = e2_val.upper().replace(' ', '')
            if norm == '=D2':
                e_correct += 1

        # Check E3:E51 = =E{row-1}+D{row}
        for row in range(3, 52):
            val = ws.cell(row=row, column=5).value
            if val is not None and isinstance(val, str):
                norm = val.upper().replace(' ', '')
                expected = f'=E{row-1}+D{row}'
                expected_norm = expected.upper().replace(' ', '')
                if norm == expected_norm:
                    e_correct += 1

        ratio = e_correct / e_total
        if ratio == 1.0:
            print(f"PASS: Component 2 — Cumulative % formulas: all {e_correct}/{e_total} rows correct (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.8:
            partial = round(0.25 * ratio, 4)
            print(f"PARTIAL: Component 2 — Cumulative % formulas: {e_correct}/{e_total} rows correct (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Cumulative % formulas: only {e_correct}/{e_total} rows correct")
            print(f"  E2={repr(ws.cell(2, 5).value)}, E3={repr(ws.cell(3, 5).value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Cumulative % formulas: {e}")

    # Component 3: ABC class IFS formulas in F2:F51 (0.25 points)
    # Each should be an IFS formula classifying A (<=0.80), B (<=0.95), C (otherwise)
    try:
        f_correct = 0
        f_total = 50  # rows 2-51
        for row in range(2, 52):
            val = ws.cell(row=row, column=6).value
            if val is not None and isinstance(val, str):
                norm = val.upper().replace(' ', '')
                # Must be IFS formula referencing the E column cumulative %
                # Accept any IFS formula that references E{row} with 0.8 and 0.95 thresholds
                if (norm.startswith('=IFS') and
                        f'E{row}' in val.upper() and
                        ('0.8' in norm or '0.80' in norm) and
                        ('0.95' in norm)):
                    f_correct += 1

        ratio = f_correct / f_total
        if ratio == 1.0:
            print(f"PASS: Component 3 — ABC IFS formulas: all {f_correct}/{f_total} rows correct (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.8:
            partial = round(0.25 * ratio, 4)
            print(f"PARTIAL: Component 3 — ABC IFS formulas: {f_correct}/{f_total} rows correct (partial {partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — ABC IFS formulas: only {f_correct}/{f_total} rows correct")
            print(f"  F2={repr(ws.cell(2, 6).value)}")
    except Exception as e:
        print(f"ERROR: Component 3 — ABC IFS formulas: {e}")

    # Component 4: Pie chart present with SUMIF source data in H/I columns (0.25 points)
    # Sub-check A: A PieChart exists on the Products sheet (0.15 pts)
    # Sub-check B: SUMIF formulas for A, B, C categories are in column I (0.10 pts)
    try:
        charts = ws._charts
        # Count pie charts using actual API type check
        pie_count = sum(1 for c in charts if type(c).__name__ == 'PieChart')

        if pie_count >= 1:
            print(f"PASS: Component 4a — Pie chart found on Products sheet ({pie_count} pie chart(s)) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4a — No pie chart found on Products sheet (expected a PieChart, got {len(charts)} chart(s))")

        # Check SUMIF formulas in I2:I4 for A, B, C respectively
        sumif_correct = 0
        for row, cls in [(2, 'A'), (3, 'B'), (4, 'C')]:
            val = ws.cell(row=row, column=9).value
            if val is not None and isinstance(val, str):
                norm = val.upper().replace(' ', '')
                # Should be a SUMIF formula referencing F column for class and C column for revenue
                if ('SUMIF' in norm and
                        f'"${cls}"' in val.upper().replace(' ', '') or
                        f'"{cls}"' in val.upper().replace(' ', '')):
                    if ('$F$2:$F$51' in val.upper().replace(' ', '') or
                            'F$2:F$51' in val.upper().replace(' ', '') or
                            'F2:F51' in val.upper().replace(' ', '')) and \
                       ('$C$2:$C$51' in val.upper().replace(' ', '') or
                            'C$2:C$51' in val.upper().replace(' ', '') or
                            'C2:C51' in val.upper().replace(' ', '')):
                        sumif_correct += 1

        if sumif_correct == 3:
            print(f"PASS: Component 4b — SUMIF formulas for A/B/C in I2:I4 (0.10 pts)")
            total_score += 0.10
        elif sumif_correct > 0:
            partial = round(0.10 * sumif_correct / 3, 4)
            print(f"PARTIAL: Component 4b — {sumif_correct}/3 SUMIF formulas correct (partial {partial} pts)")
            print(f"  I2={repr(ws.cell(2, 9).value)}, I3={repr(ws.cell(3, 9).value)}, I4={repr(ws.cell(4, 9).value)}")
            if partial > 0:
                total_score += partial
        else:
            print(f"FAIL: Component 4b — SUMIF formulas not found in I2:I4")
            print(f"  I2={repr(ws.cell(2, 9).value)}, I3={repr(ws.cell(3, 9).value)}, I4={repr(ws.cell(4, 9).value)}")
    except Exception as e:
        print(f"ERROR: Component 4 — Pie chart / SUMIF check: {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
