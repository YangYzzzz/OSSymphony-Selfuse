"""
Reward Script: Wrap MATCH formulas in C2:C11 with IFNA to show 0 for missing matches.
Task ID: calc_fma_ifna_028
Domain: libreoffice_calc
Scoring:
  Component 1: At least one cell in C2:C11 has an IFNA-wrapped MATCH formula (0.4 pts)
  Component 2: All 10 cells C2:C11 have IFNA-wrapped MATCH formulas (0.3 pts)
  Component 3: Formulas reference correct lookup range $B$14:$B$21 with exact match and 0 fallback (0.3 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fma_ifna_028'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Lookup' sheet must exist
    if 'Lookup' not in wb.sheetnames:
        print("FAIL: 'Lookup' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Lookup']

    # Collect formula values from C2:C11
    formulas = {}
    for row in range(2, 12):
        val = ws.cell(row=row, column=3).value
        formulas[row] = val

    # Component 1: At least one cell in C2:C11 has an IFNA-wrapped MATCH formula (0.4 pts)
    # This checks that the agent has started wrapping formulas with IFNA.
    # The initial file has all None values in C2:C11, so any IFNA formula indicates progress.
    try:
        ifna_match_pattern = re.compile(r'=IFNA\s*\(\s*MATCH\s*\(', re.IGNORECASE)
        cells_with_ifna = []
        for row in range(2, 12):
            val = formulas.get(row)
            if val and isinstance(val, str) and ifna_match_pattern.match(val.strip()):
                cells_with_ifna.append(f"C{row}")
        if len(cells_with_ifna) >= 1:
            print(f"PASS: Component 1 — {len(cells_with_ifna)} cell(s) have IFNA-wrapped MATCH formula: {cells_with_ifna[:3]}... (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — No IFNA-wrapped MATCH formulas found in C2:C11. Values: {list(formulas.values())}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 10 cells C2:C11 have IFNA-wrapped MATCH formulas (0.3 pts)
    # This checks that the agent completed the wrapping for all rows.
    try:
        ifna_match_pattern2 = re.compile(r'=IFNA\s*\(\s*MATCH\s*\(', re.IGNORECASE)
        cells_missing = []
        for row in range(2, 12):
            val = formulas.get(row)
            if not (val and isinstance(val, str) and ifna_match_pattern2.match(val.strip())):
                cells_missing.append(f"C{row}")
        if len(cells_missing) == 0:
            print(f"PASS: Component 2 — All 10 cells C2:C11 have IFNA-wrapped MATCH formulas (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — {len(cells_missing)} cell(s) missing IFNA-wrapped MATCH formula: {cells_missing}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formulas reference correct lookup range $B$14:$B$21 with exact match (0) and 0 fallback (0.3 pts)
    # This verifies the formula structure matches the task requirement precisely:
    # =IFNA(MATCH(An,$B$14:$B$21,0),0)
    try:
        correct_structure_pattern = re.compile(
            r'=IFNA\s*\(\s*MATCH\s*\(\s*A\d+\s*,\s*\$B\$14:\$B\$21\s*,\s*0\s*\)\s*,\s*0\s*\)',
            re.IGNORECASE
        )
        cells_correct_structure = []
        cells_wrong_structure = []
        for row in range(2, 12):
            val = formulas.get(row)
            if val and isinstance(val, str):
                if correct_structure_pattern.match(val.strip()):
                    cells_correct_structure.append(f"C{row}")
                else:
                    cells_wrong_structure.append(f"C{row} ({repr(val)})")
        if len(cells_correct_structure) == 10:
            print(f"PASS: Component 3 — All 10 formulas correctly reference $B$14:$B$21 with match_type=0 and fallback=0 (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — Only {len(cells_correct_structure)}/10 formulas have correct structure.")
            if cells_wrong_structure:
                print(f"  Cells with wrong structure: {cells_wrong_structure[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
