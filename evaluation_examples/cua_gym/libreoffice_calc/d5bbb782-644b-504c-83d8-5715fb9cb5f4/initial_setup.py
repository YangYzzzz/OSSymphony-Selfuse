"""
Initial Setup: MATCH formula wrapping with IFNA task
Task ID: calc_fma_ifna_028
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fma_ifna_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Lookup ---
    ws = wb.active
    ws.title = 'Lookup'

    # Column A header
    ws['A1'] = 'Search Value'
    ws['A1'].font = Font(bold=True)

    # Column A data (rows 2-11): search values
    search_values = [
        'Apple',
        'Banana',
        'Cherry',
        'Durian',
        'Elderberry',
        'Fig',
        'Grape',
        'Honeydew',
        'Jackfruit',
        'Kiwi',
    ]
    for i, val in enumerate(search_values, start=2):
        ws.cell(row=i, column=1, value=val)

    # Column B header (row 13 acts as a section header)
    ws['B13'] = 'Lookup List'
    ws['B13'].font = Font(bold=True)

    # Column B data (rows 14-21): the lookup list
    lookup_list = [
        'Apple',
        'Cherry',
        'Elderberry',
        'Fig',
        'Grape',
        'Kiwi',
        'Mango',
        'Papaya',
    ]
    for i, val in enumerate(lookup_list, start=14):
        ws.cell(row=i, column=2, value=val)

    # Column C header (row 1)
    ws['C1'] = 'Position'
    ws['C1'].font = Font(bold=True)

    # C2:C11 intentionally LEFT EMPTY — agent must fill these with IFNA(MATCH(...)) formulas

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
