"""
Reward Script: Unhide row 15 in sales_pipeline.xlsx
Task ID: calc_gfl_020
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): Row 15 is visible (not hidden)
  Component 2 (0.3): Row 15 data is intact (key cell values preserved)
  Component 3 (0.2): No other rows (1-30) were inadvertently hidden
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_020'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Pipeline']

    # Component 1: Row 15 is visible / not hidden (0.5 points)
    # This is the core task requirement. In initial_env row 15 is hidden;
    # in golden_env it should be visible.
    try:
        row15_hidden = ws.row_dimensions[15].hidden
        if not row15_hidden:
            print(f"PASS: Component 1 - Row 15 is visible (hidden={row15_hidden}) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 - Row 15 is still hidden (hidden={row15_hidden})")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Row 15 data integrity (0.3 points)
    # Verify that key cell values in row 15 are preserved after unhiding.
    # Expected: DEAL-015, Acme Corp, John Smith, 250000, Negotiation, Sarah K, 0.85
    # This component only awards points when COMBINED with row 15 being visible,
    # ensuring it does not score on initial_env (where row 15 is hidden).
    try:
        row15_visible = not ws.row_dimensions[15].hidden
        deal_id = ws.cell(row=15, column=1).value
        company = ws.cell(row=15, column=2).value
        value_cell = ws.cell(row=15, column=4).value
        stage = ws.cell(row=15, column=5).value

        checks_passed = 0
        if deal_id == 'DEAL-015':
            checks_passed += 1
        if company == 'Acme Corp':
            checks_passed += 1
        if value_cell == 250000 or (isinstance(value_cell, (int, float)) and abs(float(value_cell) - 250000) < 1):
            checks_passed += 1
        if stage == 'Negotiation':
            checks_passed += 1

        # Only award points if row 15 is visible AND data is intact
        if row15_visible and checks_passed == 4:
            print(f"PASS: Component 2 - Row 15 data intact: {deal_id}, {company}, {value_cell}, {stage} (0.3 pts)")
            total_score += 0.3
        elif row15_visible and checks_passed > 0:
            partial = round(0.3 * checks_passed / 4, 2)
            print(f"PARTIAL: Component 2 - {checks_passed}/4 data checks passed (row visible) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Row 15 not visible or data corrupted (visible={row15_visible}, checks={checks_passed}/4)")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: No other rows (1-14, 16-30) were inadvertently hidden (0.2 points)
    # Only score this if row 15 is visible — anchoring to the task change.
    try:
        row15_visible = not ws.row_dimensions[15].hidden
        other_hidden = []
        for r in range(1, 31):
            if r == 15:
                continue
            if ws.row_dimensions[r].hidden:
                other_hidden.append(r)

        if row15_visible and len(other_hidden) == 0:
            print(f"PASS: Component 3 - No other rows hidden, structure intact (0.2 pts)")
            total_score += 0.2
        elif not row15_visible:
            print(f"FAIL: Component 3 - Row 15 not visible, skipping other-row check")
        else:
            print(f"FAIL: Component 3 - Other rows were inadvertently hidden: {other_hidden}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
