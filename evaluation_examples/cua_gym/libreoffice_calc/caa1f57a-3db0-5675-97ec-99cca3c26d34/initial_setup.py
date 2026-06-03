"""
Initial Setup: Unhide row 15 which was previously hidden and contains missing data
Task ID: calc_gfl_020
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_020'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Pipeline'

    # Headers
    headers = ['Deal ID', 'Company', 'Contact', 'Value', 'Stage', 'Close Date', 'Owner', 'Probability']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    white_font = Font(bold=True, size=11, name='Calibri', color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = white_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Sales pipeline data (rows 2-30, i.e. 29 deals)
    deals = [
        ['DEAL-001', 'TechVision Inc', 'Emily Zhang', 45000, 'Prospecting', '2024-01-15', 'Michael R', 0.20],
        ['DEAL-002', 'Global Dynamics', 'Robert Chen', 120000, 'Qualification', '2024-02-10', 'Sarah K', 0.35],
        ['DEAL-003', 'Pinnacle Solutions', 'Amanda Torres', 78000, 'Proposal', '2024-01-28', 'David L', 0.60],
        ['DEAL-004', 'Nexus Enterprises', 'James Wilson', 95000, 'Negotiation', '2024-03-05', 'Michael R', 0.75],
        ['DEAL-005', 'Horizon Labs', 'Priya Patel', 62000, 'Closed Won', '2024-01-20', 'Sarah K', 1.00],
        ['DEAL-006', 'Summit Group', 'Carlos Rivera', 88000, 'Proposal', '2024-02-18', 'David L', 0.55],
        ['DEAL-007', 'Vertex Analytics', 'Lisa Nakamura', 34000, 'Prospecting', '2024-03-12', 'Jennifer W', 0.15],
        ['DEAL-008', 'Cascade Systems', 'Daniel Kim', 156000, 'Qualification', '2024-02-25', 'Michael R', 0.40],
        ['DEAL-009', 'Sterling Corp', 'Maria Gonzalez', 71000, 'Closed Lost', '2024-01-10', 'Sarah K', 0.00],
        ['DEAL-010', 'BluePeak Tech', 'Andrew Foster', 43000, 'Proposal', '2024-03-20', 'David L', 0.50],
        ['DEAL-011', 'Quantum Dynamics', 'Rachel Lee', 110000, 'Negotiation', '2024-02-14', 'Jennifer W', 0.80],
        ['DEAL-012', 'Ironclad Security', 'Thomas Brown', 67000, 'Qualification', '2024-03-08', 'Michael R', 0.30],
        ['DEAL-013', 'Meridian Health', 'Sandra Okafor', 89000, 'Proposal', '2024-01-30', 'Sarah K', 0.65],
        # Row 15 (index 13 in data, row 15 in sheet) - the hidden deal
        ['DEAL-015', 'Acme Corp', 'John Smith', 250000, 'Negotiation', '2024-03-31', 'Sarah K', 0.85],
        # Rows 16-30
        ['DEAL-016', 'Crestline Partners', 'Kevin Murphy', 54000, 'Prospecting', '2024-04-05', 'David L', 0.10],
        ['DEAL-017', 'Solaris Energy', 'Fatima Al-Hassan', 185000, 'Qualification', '2024-03-18', 'Jennifer W', 0.45],
        ['DEAL-018', 'Atlas Manufacturing', 'Brian O\'Neill', 73000, 'Proposal', '2024-04-12', 'Michael R', 0.55],
        ['DEAL-019', 'Vanguard Logistics', 'Yuki Tanaka', 98000, 'Closed Won', '2024-02-28', 'Sarah K', 1.00],
        ['DEAL-020', 'Ember Creative', 'Nicole Peters', 31000, 'Prospecting', '2024-04-20', 'David L', 0.20],
        ['DEAL-021', 'Cobalt Industries', 'Marcus Johnson', 142000, 'Negotiation', '2024-03-25', 'Jennifer W', 0.70],
        ['DEAL-022', 'Prism Analytics', 'Diana Reyes', 56000, 'Qualification', '2024-04-08', 'Michael R', 0.35],
        ['DEAL-023', 'Aegis Defense', 'William Chang', 210000, 'Proposal', '2024-03-15', 'Sarah K', 0.60],
        ['DEAL-024', 'Lumina Biotech', 'Olivia Bennett', 87000, 'Closed Lost', '2024-02-05', 'David L', 0.00],
        ['DEAL-025', 'Titan Consulting', 'Hassan Ibrahim', 64000, 'Prospecting', '2024-04-15', 'Jennifer W', 0.15],
        ['DEAL-026', 'Redwood Partners', 'Stephanie Liu', 105000, 'Negotiation', '2024-03-22', 'Michael R', 0.75],
        ['DEAL-027', 'Phoenix Software', 'Ryan Cooper', 48000, 'Qualification', '2024-04-18', 'Sarah K', 0.30],
        ['DEAL-028', 'Catalyst Ventures', 'Laura Martinez', 93000, 'Proposal', '2024-03-10', 'David L', 0.50],
        ['DEAL-029', 'Nordic Systems', 'Erik Johansson', 77000, 'Closed Won', '2024-01-25', 'Jennifer W', 1.00],
        ['DEAL-030', 'Silverline Media', 'Grace Adeyemi', 39000, 'Prospecting', '2024-04-22', 'Michael R', 0.10],
    ]

    # Write data
    for r, row_data in enumerate(deals, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Format Value column as currency
    for r in range(2, 31):
        ws.cell(row=r, column=4).number_format = '$#,##0'

    # Format Probability column as percentage
    for r in range(2, 31):
        ws.cell(row=r, column=8).number_format = '0%'

    # Set column widths for readability
    col_widths = {'A': 12, 'B': 22, 'C': 18, 'D': 14, 'E': 16, 'F': 14, 'G': 12, 'H': 14}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Hide row 15 (the Acme Corp deal flagged for review)
    ws.row_dimensions[15].hidden = True

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
