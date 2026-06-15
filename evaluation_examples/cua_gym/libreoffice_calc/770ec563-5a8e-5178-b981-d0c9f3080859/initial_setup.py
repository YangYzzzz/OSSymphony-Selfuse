"""
Initial Setup: Apply a 3-arrow icon set conditional formatting to trend column
Task ID: calc_fmt_condfmt_iconset_047
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_condfmt_iconset_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Trend Analysis ---
    ws = wb.active
    ws.title = 'Trend Analysis'

    # Headers
    headers = ['Metric', 'Previous', 'Current', 'Trend']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Data rows (rows 2-20, 19 rows of realistic business metrics)
    # Column D (Trend) contains values from -50 to +50; NO conditional formatting
    data = [
        ('Revenue Growth',       4820500, 5134200,   31),
        ('Customer Retention',     88.4,    91.2,    27),
        ('Operating Margin',       22.1,    19.8,   -23),
        ('Employee Satisfaction',  7.2,     7.2,     0),
        ('Net Promoter Score',     42,      47,      5),
        ('Market Share',           15.3,    14.8,   -5),
        ('Product Returns',        3.2,     2.8,   -40),
        ('Website Conversion',     3.8,     4.6,    42),
        ('Support Ticket Volume',  1240,    1180,  -48),
        ('Average Order Value',    127.50,  138.00,  10),
        ('Churn Rate',             4.5,     4.5,     0),
        ('Pipeline Coverage',      3.1,     2.6,   -16),
        ('New Customer Acq.',      312,     389,     25),
        ('Time to Resolution',     4.2,     3.8,    -9),
        ('Gross Profit Margin',    58.7,    61.2,    43),
        ('Inventory Turnover',     8.4,     7.9,    -6),
        ('Social Media Reach',     82400,   95100,   15),
        ('Subscription Renewals',  76.8,    76.8,    0),
        ('Training Completion',    68,      79,      50),
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Metric name
        ws.cell(row=r, column=2, value=row_data[1])  # Previous
        ws.cell(row=r, column=3, value=row_data[2])  # Current
        ws.cell(row=r, column=4, value=row_data[3])  # Trend (-50 to +50)

    # Column widths
    ws.column_dimensions['A'].width = 26
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10

    # Freeze top row
    ws.freeze_panes = 'A2'

    # NOTE: NO conditional formatting on D2:D20 in the initial file
    # The task requires the agent to ADD the icon set conditional formatting

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: Trend Analysis')
    print(f'Data rows: 19 (rows 2-20)')
    print(f'Column D (Trend) values: range from -50 to +50')
    print(f'Conditional formatting: NONE (task requires agent to add it)')


create_initial()
