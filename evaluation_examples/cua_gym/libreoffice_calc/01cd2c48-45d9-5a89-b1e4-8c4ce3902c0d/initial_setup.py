"""
Initial Setup: Create spreadsheet with employee data where some rows have empty column A.
Task ID: calc_mcp_009
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'CleanUp'

    # --- Headers ---
    headers = ['Employee ID', 'Full Name', 'Department', 'Salary', 'Start Date']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Realistic employee data ---
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'David', 'Mei',
        'Carlos', 'Fatima', 'Robert', 'Yuki', 'Thomas', 'Amara', 'Kevin',
        'Sofia', 'Michael', 'Lena', 'Ahmed', 'Julia', 'Raj', 'Olivia',
        'Daniel', 'Naomi', 'Ryan', 'Isabella', 'Wei', 'Grace', 'Patrick',
        'Hannah', 'Leo', 'Chloe', 'Nathan', 'Zara', 'Owen', 'Maya',
        'Victor', 'Aria', 'Samuel', 'Nina', 'Felix', 'Emma', 'Andre',
        'Lily', 'Jordan', 'Clara', 'Sean', 'Vera', 'Ian', 'Rosa', 'Derek',
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Williams', 'Sharma', 'Kim', 'Zhang',
        'Rodriguez', 'Ali', 'Taylor', 'Tanaka', 'Brown', 'Okafor', 'Lee',
        'Martinez', 'Davis', 'Muller', 'Hassan', 'Fischer', 'Patel', 'Clark',
        'Nguyen', 'Sato', 'Murphy', 'Lopez', 'Wang', 'Thompson', 'O\'Brien',
        'Miller', 'Rossi', 'Dubois', 'Scott', 'Khan', 'Evans', 'Singh',
        'Ivanov', 'Park', 'Adams', 'Yamamoto', 'Berg', 'Wright', 'Moreau',
        'Torres', 'Baker', 'Weber', 'Reyes', 'Kowalski', 'Green', 'Pereira', 'Hall',
    ]
    departments = [
        'Engineering', 'Marketing', 'Finance', 'Human Resources', 'Sales',
        'Operations', 'Research', 'Customer Support', 'Legal', 'Product',
    ]
    start_years = list(range(2018, 2026))
    months = list(range(1, 13))

    # Rows that will have empty column A (scattered throughout rows 2-200)
    # We'll create 55 rows total, with about 12 having empty column A
    empty_a_rows = {5, 11, 18, 23, 31, 37, 42, 48, 56, 63, 71, 78, 85, 92, 104}

    total_data_rows = 120
    emp_id_counter = 1001

    for r in range(2, 2 + total_data_rows):
        row_idx = r
        name_idx = (r - 2) % len(first_names)
        lname_idx = (r - 2 + 7) % len(last_names)

        full_name = f'{first_names[name_idx]} {last_names[lname_idx]}'
        dept = departments[random.randint(0, len(departments) - 1)]
        salary = round(random.uniform(45000, 135000), 2)
        year = random.choice(start_years)
        month = random.choice(months)
        day = random.randint(1, 28)
        start_date = f'{year}-{month:02d}-{day:02d}'

        if r in empty_a_rows:
            # Leave column A empty for these rows
            ws.cell(row=row_idx, column=1, value=None)
        else:
            ws.cell(row=row_idx, column=1, value=f'EMP-{emp_id_counter}')
            emp_id_counter += 1

        ws.cell(row=row_idx, column=2, value=full_name)
        ws.cell(row=row_idx, column=3, value=dept)
        ws.cell(row=row_idx, column=4, value=salary)
        ws.cell(row=row_idx, column=4).number_format = '$#,##0.00'
        ws.cell(row=row_idx, column=5, value=start_date)

    # Column widths for readability
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
