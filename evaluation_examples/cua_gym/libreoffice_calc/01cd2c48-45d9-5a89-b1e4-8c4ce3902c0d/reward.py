"""
Reward Script: Delete empty rows in column A via macro
Task ID: calc_mcp_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): No empty cells in column A from row 2 onward
  Component 2 (0.3): Row count is 106 (header + 105 data rows, empty rows deleted not just cleared)
  Component 3 (0.3): All 105 original employee IDs preserved in correct order
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_009'

# The 105 expected employee IDs (from initial non-empty rows, in order)
EXPECTED_IDS = [f'EMP-{i}' for i in range(1001, 1106)]


def persist_app_state(domain: str):
    """Save any unsaved LibreOffice changes before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'CleanUp' sheet exists (precondition gate)
    if 'CleanUp' not in wb.sheetnames:
        print("CRITICAL: 'CleanUp' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CleanUp']

    # Component 1: No empty cells in column A from row 2 onward (0.4 points)
    # This FAILS on initial (15 empty rows) and PASSES on golden (0 empty rows)
    try:
        empty_count = 0
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is None or str(val).strip() == '':
                empty_count += 1
        if empty_count == 0:
            print(f"PASS: Component 1 — No empty cells in column A (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — Found {empty_count} empty cells in column A")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row count is exactly 106 (header + 105 data rows) (0.3 points)
    # Initial has 121 rows (15 empty + 105 data + 1 header). Golden has 106.
    # This ensures rows were actually DELETED (shifted up), not just cleared.
    try:
        max_row = ws.max_row
        if max_row == 106:
            print(f"PASS: Component 2 — Row count is 106 (header + 105 data) (0.3 pts)")
            total_score += 0.3
        elif max_row <= 106:
            # Close but not exact — partial credit if within range and no data lost
            print(f"PARTIAL: Component 2 — Row count is {max_row}, expected 106 (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 2 — Row count is {max_row}, expected 106 (rows not fully deleted)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 105 employee IDs in contiguous rows 2-106 with correct order (0.3 points)
    # Initial has IDs scattered with empty-row gaps; golden has them packed in rows 2-106.
    # We check that rows 2 through 106 each have the expected ID — this FAILS on initial
    # because row 5 is empty (and subsequent IDs are shifted).
    try:
        mismatch_count = 0
        mismatch_details = []
        for i, expected_id in enumerate(EXPECTED_IDS):
            r = i + 2  # row 2 = first data row
            val = ws.cell(row=r, column=1).value
            actual_str = str(val).strip() if val is not None else ''
            if actual_str != expected_id:
                mismatch_count += 1
                if len(mismatch_details) < 3:
                    mismatch_details.append(f"Row {r}: expected '{expected_id}', found '{actual_str}'")

        if mismatch_count == 0:
            print(f"PASS: Component 3 — All 105 IDs in contiguous rows 2-106 in correct order (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — IDs not contiguous from row 2. First mismatches: {mismatch_details}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
