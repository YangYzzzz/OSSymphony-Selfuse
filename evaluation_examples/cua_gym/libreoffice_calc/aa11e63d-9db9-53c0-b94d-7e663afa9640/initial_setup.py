"""
Initial Setup: Insert a blank row above row 5 to add visual separation before the totals section.
Task ID: calc_cop_insert_row_col_001
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_insert_row_col_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Expenses ---
    ws = wb.active
    ws.title = 'Expenses'

    # Headers (row 1)
    headers = ['Category', 'Q1', 'Q2', 'Q3', 'Q4', 'Annual Total']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF')

    # Data rows (rows 2-4): expense categories with realistic quarterly data
    data = [
        ['Office Supplies',    3420.50,  2980.75,  3150.00,  4200.25,  '=SUM(B2:E2)'],
        ['Travel & Lodging',  15600.00, 12400.50, 18900.75, 21300.00,  '=SUM(B3:E3)'],
        ['Software Licenses',  8750.00,  8750.00,  9500.00,  9500.00,  '=SUM(B4:E4)'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Row 5: Totals row with SUM formulas
    ws.cell(row=5, column=1, value='Total').font = Font(bold=True)
    totals_formulas = ['=SUM(B2:B4)', '=SUM(C2:C4)', '=SUM(D2:D4)', '=SUM(E2:E4)', '=SUM(F2:F4)']
    for col_idx, formula in enumerate(totals_formulas, 2):
        cell = ws.cell(row=5, column=col_idx, value=formula)
        cell.font = Font(bold=True)

    # Column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
