"""
Reward Script: Insert a blank row above row 5 to add visual separation before the totals section.
Task ID: calc_cop_insert_row_col_001
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Row 5 in 'Expenses' sheet is blank (the newly inserted separator row)
  Component 2 (0.3): Totals row moved to row 6 — 'Total' label in A6 and SUM formulas in B6:F6
  Component 3 (0.3): SUM formulas in row 6 reference the correct data range (rows 2–4)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_insert_row_col_001'
SHEET_NAME = 'Expenses'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Insert a blank row above the original row 5 (totals row) to add visual
    separation. After insertion:
      - New row 5 should be completely blank
      - Old row 5 (totals) should now be at row 6
      - SUM formulas in row 6 should still reference data rows 2-4
    """
    total_score = 0.0

    # Load workbook — failure is fatal
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in workbook. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # ------------------------------------------------------------------
    # Component 1: Row 5 is blank (0.4 points)
    # The task inserts a blank row at row 5 for visual separation.
    # In the initial file, row 5 contains totals. In the golden file,
    # row 5 must be completely empty (all 6 columns None).
    # This FAILS on initial (row 5 has data) → PASSES on golden (row 5 blank).
    # ------------------------------------------------------------------
    try:
        row5_values = [ws.cell(row=5, column=c).value for c in range(1, 7)]
        all_blank = all(v is None for v in row5_values)
        if all_blank:
            print(f"PASS: Component 1 — Row 5 is completely blank (all 6 columns empty) (0.4 pts)")
            total_score += 0.4
        else:
            non_blank = [(c, row5_values[c-1]) for c in range(1, 7) if row5_values[c-1] is not None]
            print(f"FAIL: Component 1 — Row 5 is NOT blank. Non-empty cells: {non_blank}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Totals row has moved to row 6 (0.3 points)
    # The original totals row (row 5 in initial) should now be at row 6.
    # Checks that A6 == 'Total' (case-insensitive) and that row 6 has
    # SUM formula strings in columns B through F.
    # This FAILS on initial (row 6 is empty) → PASSES on golden (totals at row 6).
    # ------------------------------------------------------------------
    try:
        a6 = ws.cell(row=6, column=1).value
        label_ok = a6 is not None and str(a6).strip().lower() == 'total'

        # Check that B6:F6 all contain SUM formulas
        formula_cols = []
        for col in range(2, 7):  # columns B through F
            val = ws.cell(row=6, column=col).value
            if val is not None and isinstance(val, str) and val.upper().startswith('=SUM('):
                formula_cols.append(col)

        totals_row_present = label_ok and len(formula_cols) == 5

        if totals_row_present:
            print(f"PASS: Component 2 — Totals row at row 6: A6='{a6}', SUM formulas in B6:F6 (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 2 — Expected totals at row 6. "
                  f"A6={repr(a6)}, SUM formula columns found: {formula_cols} (need 5: B-F)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: SUM formulas in row 6 reference the correct range (0.3 points)
    # The original SUM formulas in row 5 (initial) referenced rows 2-4.
    # After inserting row 5 (blank), the new row 6 formulas should STILL
    # reference rows 2-4 (not rows 2-5, which would include the blank row).
    #
    # Expected formulas:
    #   B6: =SUM(B2:B4)
    #   C6: =SUM(C2:C4)
    #   D6: =SUM(D2:D4)
    #   E6: =SUM(E2:E4)
    #   F6: =SUM(F2:F4)  (column totals)
    # F6 may also be =SUM(F2:F4) or =SUM(B6:E6) — either is acceptable.
    #
    # Key requirement: column SUM formulas (B6:E6) must reference rows 2-4,
    # which means they exclude the blank row 5 (correct behavior).
    # This FAILS on initial (row 6 doesn't exist as totals) → PASSES on golden.
    # ------------------------------------------------------------------
    try:
        correct_range_count = 0
        for col_idx, col_letter in enumerate(['B', 'C', 'D', 'E'], start=2):
            formula = ws.cell(row=6, column=col_idx).value
            if formula is not None and isinstance(formula, str):
                # Normalize: remove spaces, uppercase
                norm = formula.upper().replace(' ', '')
                # Correct: references rows 2-4 for this column
                expected = f'=SUM({col_letter}2:{col_letter}4)'
                if norm == expected:
                    correct_range_count += 1
                else:
                    print(f"  INFO: {col_letter}6 formula = {repr(formula)} (expected {expected})")

        # All 4 quarter-column formulas must reference rows 2-4
        if correct_range_count == 4:
            print(f"PASS: Component 3 — SUM formulas in B6:E6 correctly reference rows 2-4 (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 3 — Only {correct_range_count}/4 column formulas reference rows 2-4 correctly")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
