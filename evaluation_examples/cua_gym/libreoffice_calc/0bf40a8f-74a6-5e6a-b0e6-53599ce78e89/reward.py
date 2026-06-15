"""
Reward Script: Warehouse Space Utilization Report
Task ID: calc_wf_066
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25) - Utilization % formulas in E4:E11
  Component 2 (0.25) - Status/rebalance formulas in F4:F11
  Component 3 (0.15) - Conditional formatting on utilization column (3 rules)
  Component 4 (0.15) - Bar chart present for utilization by zone
  Component 5 (0.20) - Summary statistics section (Total, Avg, Min, Max)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_066'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify warehouse space utilization report task completion.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Warehouse' sheet must exist
    if 'Warehouse' not in wb.sheetnames:
        print("CRITICAL: 'Warehouse' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Warehouse']

    # Component 1: Utilization % formulas in E4:E11 (0.25 points)
    # Each cell should have a formula like =D4/C4*100 (or equivalent)
    # Initial state: E4:E11 are empty. Golden: formulas present.
    try:
        util_formula_count = 0
        for row in range(4, 12):
            cell_val = ws.cell(row=row, column=5).value  # Column E
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(" ", "")
                # Check for utilization formula pattern: references D and C of same row
                if (f"D{row}" in val_upper and f"C{row}" in val_upper and
                        ("/" in val_upper or "*" in val_upper)):
                    util_formula_count += 1
        if util_formula_count == 8:
            print(f"PASS: Component 1 — All 8 utilization formulas found in E4:E11 (0.25 pts)")
            total_score += 0.25
        elif util_formula_count > 0:
            partial = round(0.25 * util_formula_count / 8, 3)
            print(f"PARTIAL: Component 1 — {util_formula_count}/8 utilization formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No utilization formulas found in E4:E11")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Status/rebalance formulas in F4:F11 (0.25 points)
    # Each cell should have IF(OR(...)) formula checking >90 or <30 for REBALANCE/OK
    # Initial state: F4:F11 are empty. Golden: formulas present.
    try:
        status_formula_count = 0
        for row in range(4, 12):
            cell_val = ws.cell(row=row, column=6).value  # Column F
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(" ", "")
                # Check for IF/OR pattern with REBALANCE and OK
                if ("IF(" in val_upper and "OR(" in val_upper and
                        "REBALANCE" in val_upper and "OK" in val_upper):
                    status_formula_count += 1
        if status_formula_count == 8:
            print(f"PASS: Component 2 — All 8 status formulas found in F4:F11 (0.25 pts)")
            total_score += 0.25
        elif status_formula_count > 0:
            partial = round(0.25 * status_formula_count / 8, 3)
            print(f"PARTIAL: Component 2 — {status_formula_count}/8 status formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No status formulas found in F4:F11")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on utilization column (0.15 points)
    # Should have 3 rules on E4:E11: green <70, yellow 70-90, red >90
    # Initial state: 0 conditional formatting rules. Golden: 3 rules.
    try:
        cf_rules_on_e = []
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if the range covers the E column utilization cells
            if 'E' in cf_range:
                for rule in cf.rules:
                    cf_rules_on_e.append(rule)

        if len(cf_rules_on_e) >= 3:
            # Verify we have the three expected rule types
            operators_found = set()
            for rule in cf_rules_on_e:
                if rule.operator:
                    operators_found.add(rule.operator)
            # Expect lessThan, between, greaterThan
            expected_ops = {'lessThan', 'between', 'greaterThan'}
            if expected_ops.issubset(operators_found):
                print(f"PASS: Component 3 — 3 conditional formatting rules with correct operators (0.15 pts)")
                total_score += 0.15
            else:
                print(f"PARTIAL: Component 3 — {len(cf_rules_on_e)} rules found but operators {operators_found} missing some of {expected_ops} (0.08 pts)")
                total_score += 0.08
        elif len(cf_rules_on_e) > 0:
            partial = round(0.15 * len(cf_rules_on_e) / 3, 3)
            print(f"PARTIAL: Component 3 — {len(cf_rules_on_e)}/3 conditional formatting rules ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found on utilization column")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Bar chart present for utilization by zone (0.15 points)
    # Initial state: 0 charts. Golden: 1 bar chart.
    try:
        charts = ws._charts
        if len(charts) >= 1:
            # Check that at least one chart is a BarChart
            bar_chart_found = False
            for chart in charts:
                chart_class = chart.__class__.__name__
                if 'Bar' in chart_class:
                    bar_chart_found = True
                    break

            if bar_chart_found:
                print(f"PASS: Component 4 — Bar chart found on Warehouse sheet (0.15 pts)")
                total_score += 0.15
            else:
                # Any chart type is partial credit
                print(f"PARTIAL: Component 4 — Chart found but not BarChart type (0.08 pts)")
                total_score += 0.08
        else:
            print(f"FAIL: Component 4 — No charts found on Warehouse sheet")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Summary statistics section (0.20 points)
    # Should have summary rows with Total, Average, Min, Max utilization formulas
    # Initial state: no summary section. Golden: rows 13-17 with formulas.
    try:
        summary_score = 0.0
        summary_items = 0

        # Check for summary section existence (look in rows 12-20 for summary labels)
        summary_found = False
        summary_start_row = None
        for row in range(12, 21):
            cell_val = ws.cell(row=row, column=1).value
            if cell_val and 'summary' in str(cell_val).lower():
                summary_found = True
                summary_start_row = row
                break

        if not summary_found:
            print(f"FAIL: Component 5 — No summary section found")
        else:
            # Look for Total/Average/Min/Max utilization in subsequent rows
            keywords = {
                'total': False,
                'average': False,
                'min': False,
                'max': False
            }
            formula_keywords = {
                'total': 'SUM',
                'average': 'AVERAGE',
                'min': 'MIN',
                'max': 'MAX'
            }

            for row in range(summary_start_row, min(summary_start_row + 8, ws.max_row + 1)):
                label = ws.cell(row=row, column=1).value
                formula_cell = ws.cell(row=row, column=2).value
                if label is None:
                    continue
                label_lower = str(label).lower()
                for key in keywords:
                    if key in label_lower:
                        # Check if corresponding formula exists
                        if formula_cell is not None and isinstance(formula_cell, str):
                            formula_upper = str(formula_cell).upper().replace(" ", "")
                            expected_func = formula_keywords[key]
                            if expected_func in formula_upper:
                                keywords[key] = True
                                summary_items += 1

            if summary_items == 4:
                print(f"PASS: Component 5 — All 4 summary statistics found with correct formulas (0.20 pts)")
                summary_score = 0.20
            elif summary_items > 0:
                summary_score = round(0.20 * summary_items / 4, 3)
                print(f"PARTIAL: Component 5 — {summary_items}/4 summary statistics found ({summary_score} pts)")
            else:
                print(f"FAIL: Component 5 — Summary section found but no statistic formulas")

            total_score += summary_score
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
