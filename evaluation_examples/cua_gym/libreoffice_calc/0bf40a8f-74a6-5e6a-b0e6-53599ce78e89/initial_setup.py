"""
Initial Setup: Warehouse space utilization report - pre-task state
Task ID: calc_wf_066
Domain: libreoffice_calc

Creates a workbook with 8 warehouse zones (A-H), max capacity, current count,
and product mix. Utilization %, status, conditional formatting, charts, and
summary are left for the agent to complete.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_066'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Warehouse ---
    ws = wb.active
    ws.title = 'Warehouse'

    # Title row
    ws.merge_cells('A1:F1')
    ws['A1'] = 'Warehouse Space Utilization Report'
    ws['A1'].font = Font(name='Arial', size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Blank row 2

    # Headers in row 3
    headers = ['Zone', 'Product Mix', 'Max Capacity (pallets)', 'Current Count', 'Utilization %', 'Status']
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Zone data: (Zone, Product Mix, Max Capacity, Current Count)
    # Utilization % and Status are intentionally left blank for the agent
    zone_data = [
        ('Zone A', 'Electronics, Small Appliances', 450, 380),
        ('Zone B', 'Furniture, Home Decor', 600, 520),
        ('Zone C', 'Food & Beverages (Cold Storage)', 300, 85),
        ('Zone D', 'Clothing & Textiles', 500, 475),
        ('Zone E', 'Automotive Parts, Tools', 350, 210),
        ('Zone F', 'Paper Products, Office Supplies', 400, 365),
        ('Zone G', 'Chemicals, Cleaning Supplies', 250, 235),
        ('Zone H', 'Seasonal & Promotional Items', 550, 140),
    ]

    data_font = Font(name='Arial', size=11)
    data_align = Alignment(horizontal='center', vertical='center')
    product_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for r, (zone, products, max_cap, current) in enumerate(zone_data, 4):
        ws.cell(row=r, column=1, value=zone).font = data_font
        ws.cell(row=r, column=1).alignment = data_align
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2, value=products).font = data_font
        ws.cell(row=r, column=2).alignment = product_align
        ws.cell(row=r, column=2).border = thin_border

        ws.cell(row=r, column=3, value=max_cap).font = data_font
        ws.cell(row=r, column=3).alignment = data_align
        ws.cell(row=r, column=3).border = thin_border
        ws.cell(row=r, column=3).number_format = '#,##0'

        ws.cell(row=r, column=4, value=current).font = data_font
        ws.cell(row=r, column=4).alignment = data_align
        ws.cell(row=r, column=4).border = thin_border
        ws.cell(row=r, column=4).number_format = '#,##0'

        # Utilization % (column 5) - LEFT EMPTY for agent
        ws.cell(row=r, column=5).border = thin_border
        ws.cell(row=r, column=5).alignment = data_align

        # Status (column 6) - LEFT EMPTY for agent
        ws.cell(row=r, column=6).border = thin_border
        ws.cell(row=r, column=6).alignment = data_align

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 14

    # Row heights
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
