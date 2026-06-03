"""
Reward Script: School canteen daily sales report in LibreOffice Calc
Task ID: calc_grs_025
Domain: libreoffice_calc
Scoring:
  Component 1: Revenue section with formulas (0.25)
  Component 2: Cost section with formulas (0.20)
  Component 3: Profit section with formulas (0.15)
  Component 4: Summary rows/columns (daily totals + monthly totals) (0.15)
  Component 5: Conditional formatting on daily revenue row (0.10)
  Component 6: Monetary number formatting ($#,##0.00) (0.05)
  Component 7: Line chart on Charts sheet (0.10)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_025'


def find_section_row(ws, label_keyword, max_row=60):
    """Find the row where a section label starts (e.g., 'REVENUE', 'COST', 'PROFIT')."""
    for r in range(1, max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(val, str) and label_keyword.upper() in val.upper():
            return r
    return None


def has_multiplication_formula(formula_str, qty_ref_pattern, price_col_pattern):
    """Check if a formula multiplies quantity cell by a Products sheet price/cost cell."""
    if not isinstance(formula_str, str):
        return False
    f = formula_str.upper().replace(' ', '')
    # Should reference both a local cell (qty) and Products! sheet
    return 'PRODUCTS!' in f and '*' in f


def has_subtraction_formula(formula_str):
    """Check if a formula subtracts (revenue - cost)."""
    if not isinstance(formula_str, str):
        return False
    return '-' in formula_str and '=' in formula_str


def has_sum_formula(formula_str):
    """Check if a formula contains SUM."""
    if not isinstance(formula_str, str):
        return False
    return 'SUM' in formula_str.upper()


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Must have Daily Sales sheet
    if 'Daily Sales' not in wb.sheetnames:
        print("FAIL: 'Daily Sales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Daily Sales']

    # =========================================================================
    # Component 1: Revenue section with formulas (0.25 points)
    # Task: Add formulas to calculate Revenue per product per day
    # Golden: Rows 12-21 have REVENUE header, product rows with =qty*Products!D_ formulas
    # Initial: No revenue section exists
    # =========================================================================
    try:
        rev_row = find_section_row(ws, 'REVENUE')
        if rev_row is not None:
            # Check that product rows below the header have multiplication formulas
            # referencing Products!D (Price column)
            rev_formula_count = 0
            expected_product_rows = 8  # 8 products
            for offset in range(1, expected_product_rows + 1):
                r = rev_row + offset
                # Check a few date columns for revenue formulas
                for c in [3, 6, 8]:  # columns C, F, H (sample date columns)
                    cell_val = ws.cell(row=r, column=c).value
                    if has_multiplication_formula(str(cell_val) if cell_val else '', '', ''):
                        rev_formula_count += 1
                        break  # one per row is enough

            if rev_formula_count >= 6:
                print(f"PASS: Component 1 - Revenue section found at row {rev_row} with {rev_formula_count}/8 product formula rows (0.25 pts)")
                total_score += 0.25
            elif rev_formula_count >= 3:
                partial = 0.15
                print(f"PARTIAL: Component 1 - Revenue section found but only {rev_formula_count}/8 product rows have formulas ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 1 - Revenue section found at row {rev_row} but insufficient formula rows ({rev_formula_count})")
        else:
            print("FAIL: Component 1 - No 'REVENUE' section label found in column A")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Cost section with formulas (0.20 points)
    # Task: Add formulas to calculate Cost per product per day
    # Golden: Rows 23-32 have COST header, product rows with =qty*Products!E_ formulas
    # Initial: No cost section exists
    # =========================================================================
    try:
        cost_row = find_section_row(ws, 'COST')
        if cost_row is not None:
            cost_formula_count = 0
            for offset in range(1, expected_product_rows + 1):
                r = cost_row + offset
                for c in [3, 6, 8]:
                    cell_val = ws.cell(row=r, column=c).value
                    if has_multiplication_formula(str(cell_val) if cell_val else '', '', ''):
                        cost_formula_count += 1
                        break

            if cost_formula_count >= 6:
                print(f"PASS: Component 2 - Cost section found at row {cost_row} with {cost_formula_count}/8 product formula rows (0.20 pts)")
                total_score += 0.20
            elif cost_formula_count >= 3:
                partial = 0.12
                print(f"PARTIAL: Component 2 - Cost section found but only {cost_formula_count}/8 rows ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 2 - Cost section found but insufficient formula rows ({cost_formula_count})")
        else:
            print("FAIL: Component 2 - No 'COST' section label found in column A")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Profit section with formulas (0.15 points)
    # Task: Add formulas to calculate Profit per product per day (Revenue - Cost)
    # Golden: Rows 34-43 have PROFIT header, product rows with =Revenue-Cost formulas
    # Initial: No profit section exists
    # =========================================================================
    try:
        profit_row = find_section_row(ws, 'PROFIT')
        if profit_row is not None:
            profit_formula_count = 0
            for offset in range(1, expected_product_rows + 1):
                r = profit_row + offset
                for c in [3, 6, 8]:
                    cell_val = ws.cell(row=r, column=c).value
                    if has_subtraction_formula(str(cell_val) if cell_val else ''):
                        profit_formula_count += 1
                        break

            if profit_formula_count >= 6:
                print(f"PASS: Component 3 - Profit section found at row {profit_row} with {profit_formula_count}/8 product formula rows (0.15 pts)")
                total_score += 0.15
            elif profit_formula_count >= 3:
                partial = 0.08
                print(f"PARTIAL: Component 3 - Profit section only {profit_formula_count}/8 rows ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 3 - Profit section found but insufficient formula rows ({profit_formula_count})")
        else:
            print("FAIL: Component 3 - No 'PROFIT' section label found in column A")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =========================================================================
    # Component 4: Summary rows and monthly total columns (0.15 points)
    # Task: Summary row at bottom for daily totals + summary column on right for monthly totals
    # Golden: Daily Total rows with SUM formulas + AH column with SUM for monthly totals
    # Initial: No summary rows or monthly column
    # =========================================================================
    try:
        summary_score = 0.0

        # 4a: Check for daily total rows with SUM formulas (0.075 pts)
        # Look for a row labeled "Daily Total" in any section
        daily_total_found = False
        for r in range(1, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val and isinstance(val, str) and 'daily total' in val.lower():
                # Check that it has SUM formulas in date columns
                sum_count = 0
                for c in [3, 6, 8, 10]:
                    cv = ws.cell(row=r, column=c).value
                    if has_sum_formula(str(cv) if cv else ''):
                        sum_count += 1
                if sum_count >= 2:
                    daily_total_found = True
                    break

        if daily_total_found:
            print(f"PASS: Component 4a - Daily total summary row found with SUM formulas (0.075 pts)")
            summary_score += 0.075
        else:
            print("FAIL: Component 4a - No daily total summary row with SUM formulas found")

        # 4b: Check for monthly total column (rightmost col with SUM formulas) (0.075 pts)
        # Golden has AH column (col 34) with "Monthly Qty", "Monthly Revenue", etc.
        monthly_col_found = False
        # Check last few columns for SUM formulas
        for c in range(ws.max_column, max(ws.max_column - 5, 2), -1):
            header_val = ws.cell(row=1, column=c).value
            if header_val and isinstance(header_val, str) and 'monthly' in header_val.lower():
                # Check a product row in this column for SUM
                for r in [2, 3]:
                    cv = ws.cell(row=r, column=c).value
                    if has_sum_formula(str(cv) if cv else ''):
                        monthly_col_found = True
                        break
            if monthly_col_found:
                break

        # Also check if any rightmost column beyond the 31 date columns has SUM
        if not monthly_col_found:
            for c in range(ws.max_column, max(ws.max_column - 5, 2), -1):
                for r in [2, 3, 4]:
                    cv = ws.cell(row=r, column=c).value
                    if has_sum_formula(str(cv) if cv else ''):
                        monthly_col_found = True
                        break
                if monthly_col_found:
                    break

        if monthly_col_found:
            print(f"PASS: Component 4b - Monthly total column found with SUM formulas (0.075 pts)")
            summary_score += 0.075
        else:
            print("FAIL: Component 4b - No monthly total column with SUM formulas found")

        if summary_score > 0:
            total_score += summary_score
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =========================================================================
    # Component 5: Conditional formatting on daily revenue totals (0.10 points)
    # Task: Highlight days where total revenue exceeds $500 in green
    # Golden: CellIs rule on C21:AG21, greaterThan 500, green fill (FF00B050)
    # Initial: No conditional formatting
    # =========================================================================
    try:
        cf_found = False
        cf_rules = ws.conditional_formatting
        for cf in cf_rules:
            for rule in cf.rules:
                # Check if it's a "greaterThan" or "greaterThanOrEqual" type with 500
                is_cell_rule = rule.type == 'cellIs'
                is_formula_rule = rule.type == 'expression'

                formula_has_500 = False
                if rule.formula:
                    for f in rule.formula:
                        if '500' in str(f):
                            formula_has_500 = True

                has_green = False
                if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                    fg = rule.dxf.fill.fgColor
                    if fg and fg.rgb:
                        rgb = fg.rgb.upper()
                        # Check for green-ish color (00B050, 00FF00, 92D050, etc.)
                        if '00B050' in rgb or '00FF00' in rgb or '92D050' in rgb or 'GREEN' in rgb:
                            has_green = True
                        # Also accept any color with high G and low R/B
                        try:
                            r_val = int(rgb[2:4], 16)
                            g_val = int(rgb[4:6], 16)
                            b_val = int(rgb[6:8], 16)
                            if g_val > 128 and g_val > r_val and g_val > b_val:
                                has_green = True
                        except (ValueError, IndexError):
                            pass

                if (is_cell_rule or is_formula_rule) and formula_has_500:
                    if has_green:
                        cf_found = True
                        print(f"PASS: Component 5 - Conditional formatting found: revenue > 500 with green fill (0.10 pts)")
                    else:
                        # Partial: rule exists but color might not be green
                        cf_found = True
                        print(f"PASS: Component 5 - Conditional formatting found: revenue > 500 threshold (0.10 pts)")

        if cf_found:
            total_score += 0.10
        else:
            # Check if there's any conditional formatting at all (partial credit)
            if len(list(cf_rules)) > 0:
                print(f"PARTIAL: Component 5 - Conditional formatting exists but doesn't match expected criteria (0.05 pts)")
                total_score += 0.05
            else:
                print("FAIL: Component 5 - No conditional formatting found on Daily Sales sheet")
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # =========================================================================
    # Component 6: Monetary number formatting (0.05 points)
    # Task: Format all monetary values with 2 decimal places
    # Golden: Revenue/Cost/Profit cells use $#,##0.00
    # Initial: No monetary cells exist (no revenue/cost/profit sections)
    # =========================================================================
    try:
        formatted_count = 0
        # Check revenue, cost, profit cells for currency/decimal formatting
        rev_row = find_section_row(ws, 'REVENUE')
        if rev_row is not None:
            for offset in range(1, 4):  # Check first 3 product rows
                r = rev_row + offset
                nf = ws.cell(row=r, column=3).number_format
                if nf and ('0.00' in nf or '#,##0.00' in nf):
                    formatted_count += 1

        if formatted_count >= 2:
            print(f"PASS: Component 6 - Monetary values formatted with 2 decimal places (0.05 pts)")
            total_score += 0.05
        elif formatted_count >= 1:
            print(f"PARTIAL: Component 6 - Some monetary formatting found (0.03 pts)")
            total_score += 0.03
        else:
            print("FAIL: Component 6 - No monetary number formatting found on revenue/cost cells")
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # =========================================================================
    # Component 7: Line chart on Charts sheet (0.10 points)
    # Task: Create a line chart showing daily revenue trend on Sheet3
    # Golden: Charts sheet has 1 chart (line chart) titled "Daily Revenue Trend - May 2025"
    # Initial: Charts sheet has 0 charts
    # =========================================================================
    try:
        charts_sheet = None
        # Find the charts sheet (could be named 'Charts' or be the 3rd sheet)
        if 'Charts' in wb.sheetnames:
            charts_sheet = wb['Charts']
        elif len(wb.sheetnames) >= 3:
            charts_sheet = wb.worksheets[2]

        if charts_sheet is not None and len(charts_sheet._charts) >= 1:
            chart = charts_sheet._charts[0]
            print(f"PASS: Component 7 - Chart found on Charts sheet (type present, {len(charts_sheet._charts)} chart(s)) (0.10 pts)")
            total_score += 0.10
        else:
            # Check if any sheet has a chart
            chart_anywhere = False
            for sn in wb.sheetnames:
                if len(wb[sn]._charts) > 0:
                    chart_anywhere = True
                    print(f"PARTIAL: Component 7 - Chart found on sheet '{sn}' instead of Charts sheet (0.05 pts)")
                    total_score += 0.05
                    break
            if not chart_anywhere:
                print("FAIL: Component 7 - No chart found on any sheet")
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice state
def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
