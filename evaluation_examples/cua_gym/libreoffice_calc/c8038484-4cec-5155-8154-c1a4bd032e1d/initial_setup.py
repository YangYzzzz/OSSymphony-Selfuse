"""
Initial Setup: School canteen daily sales report
Task ID: calc_grs_025
Domain: libreoffice_calc

Creates:
  Sheet1 "Products" - Product master list (ID, Name, Category, Price, Cost)
  Sheet2 "Daily Sales" - Quantities sold per product per day (May 2025)
  Sheet3 "Charts" - Empty placeholder sheet

Does NOT include: formulas, summary rows/columns, conditional formatting, charts, monetary formatting.
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_025'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()

    # ===== Sheet1: Products =====
    ws1 = wb.active
    ws1.title = "Products"

    headers1 = ["Product ID", "Product Name", "Category", "Price", "Cost"]
    for col, h in enumerate(headers1, 1):
        ws1.cell(row=1, column=col, value=h)

    products = [
        ["P001", "Chicken Rice Bowl", "Main", 4.50, 2.10],
        ["P002", "Vegetable Noodle Soup", "Main", 3.80, 1.60],
        ["P003", "Fish Finger Wrap", "Main", 4.20, 1.90],
        ["P004", "Cheese Sandwich", "Snack", 2.50, 0.95],
        ["P005", "Fresh Fruit Cup", "Snack", 2.00, 0.80],
        ["P006", "Chocolate Muffin", "Snack", 1.80, 0.65],
        ["P007", "Orange Juice", "Drink", 1.50, 0.45],
        ["P008", "Milk Carton", "Drink", 1.20, 0.40],
    ]

    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws1.column_dimensions["A"].width = 12
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 10

    # ===== Sheet2: Daily Sales =====
    ws2 = wb.create_sheet("Daily Sales")

    # Row 1: Headers - Product ID, Product Name, then dates May 1-31
    ws2.cell(row=1, column=1, value="Product ID")
    ws2.cell(row=1, column=2, value="Product Name")

    num_days = 31  # May 2025
    for d in range(1, num_days + 1):
        ws2.cell(row=1, column=d + 2, value=f"May {d}")

    # Rows 2-9: One row per product with random quantities
    for i, prod in enumerate(products):
        row = i + 2
        ws2.cell(row=row, column=1, value=prod[0])  # Product ID
        ws2.cell(row=row, column=2, value=prod[1])  # Product Name

        # Generate realistic daily quantities
        if prod[2] == "Main":
            base_qty = 35
            variation = 15
        elif prod[2] == "Snack":
            base_qty = 25
            variation = 12
        else:  # Drink
            base_qty = 40
            variation = 18

        for d in range(1, num_days + 1):
            # Weekends (May 2025: Sat=3,10,17,24,31; Sun=4,11,18,25)
            day_of_week = (d + 3) % 7  # May 1, 2025 is Thursday (day_of_week 0=Mon)
            if day_of_week >= 5:
                qty = 0  # School closed on weekends
            else:
                qty = max(0, base_qty + random.randint(-variation, variation))
                # Sports day boost on May 9 and May 23
                if d in [9, 23]:
                    qty = int(qty * 1.4)
            ws2.cell(row=row, column=d + 2, value=qty)

    # Set column widths
    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 24
    for d in range(1, num_days + 1):
        col_letter = openpyxl.utils.get_column_letter(d + 2)
        ws2.column_dimensions[col_letter].width = 8

    # ===== Sheet3: Charts (empty placeholder) =====
    ws3 = wb.create_sheet("Charts")
    ws3.cell(row=1, column=1, value="Daily Revenue Trend")

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
