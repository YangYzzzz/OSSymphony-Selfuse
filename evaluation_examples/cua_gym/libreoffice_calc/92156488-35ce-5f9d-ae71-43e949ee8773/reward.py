"""
Reward Script: INDIRECT formula referencing dynamically-built sheet name
Task ID: calc_mcp_051
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.3): Dynamic!B1 contains a formula (non-empty, starts with '=')
  - Component 2 (0.3): The formula uses the INDIRECT function
  - Component 3 (0.4): The formula correctly concatenates A1 & A2 with ".C5" to
    reference Region_North!C5 — matches =INDIRECT(A1&A2&".C5") or equivalent
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_051'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice document."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that Dynamic!B1 contains an INDIRECT formula that dynamically
    references Region_North!C5 by concatenating A1 ('Region') and A2 ('_North')
    with '.C5'.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Dynamic sheet must exist
    if 'Dynamic' not in wb.sheetnames:
        print("FAIL: 'Dynamic' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Dynamic']
    b1_value = ws['B1'].value

    # Component 1: Dynamic!B1 contains a formula (0.3 points)
    # This FAILS on initial (B1 is None) and PASSES on golden (B1 has a formula)
    try:
        if b1_value is not None and isinstance(b1_value, str) and b1_value.strip().startswith('='):
            print(f"PASS: Component 1 — B1 contains a formula: {b1_value} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — B1 does not contain a formula. Value: {repr(b1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: The formula uses INDIRECT (0.3 points)
    # INDIRECT is the key function required by the task
    try:
        if b1_value and isinstance(b1_value, str):
            formula_upper = b1_value.upper().replace(" ", "")
            if 'INDIRECT(' in formula_upper:
                print(f"PASS: Component 2 — Formula uses INDIRECT function (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Formula does not use INDIRECT. Value: {repr(b1_value)}")
        else:
            print(f"FAIL: Component 2 — No formula to check for INDIRECT")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula correctly references A1&A2 concatenated with ".C5" (0.4 points)
    # The expected formula is =INDIRECT(A1&A2&".C5") or semantically equivalent forms like:
    #   =INDIRECT(A1&A2&"!C5")  — using ! separator
    #   =INDIRECT(A1&A2&".C5")  — using . separator (Calc convention)
    # Both should resolve to Region_North!C5 = 12000
    try:
        if b1_value and isinstance(b1_value, str):
            formula_norm = b1_value.upper().replace(" ", "")
            # Check that formula references A1, A2, and C5 within an INDIRECT call
            # Accept both "." and "!" as sheet-cell separators
            has_a1 = 'A1' in formula_norm
            has_a2 = 'A2' in formula_norm
            has_c5 = 'C5' in formula_norm
            has_concat = '&' in formula_norm  # string concatenation operator
            has_indirect = 'INDIRECT(' in formula_norm

            if has_indirect and has_a1 and has_a2 and has_c5 and has_concat:
                print(f"PASS: Component 3 — Formula correctly builds sheet reference from A1&A2 with C5 (0.4 pts)")
                total_score += 0.4
            else:
                missing = []
                if not has_indirect:
                    missing.append("INDIRECT()")
                if not has_a1:
                    missing.append("A1 reference")
                if not has_a2:
                    missing.append("A2 reference")
                if not has_c5:
                    missing.append("C5 reference")
                if not has_concat:
                    missing.append("& concatenation")
                print(f"FAIL: Component 3 — Formula missing: {', '.join(missing)}. Value: {repr(b1_value)}")
        else:
            print(f"FAIL: Component 3 — No formula to analyze")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
