"""
Initial Setup: Create workbook with Dynamic and Region_North sheets for INDIRECT formula task
Task ID: calc_mcp_051
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_051'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Region_North ---
    ws_rn = wb.active
    ws_rn.title = 'Region_North'

    # Headers
    headers = ['Product', 'Category', 'Revenue', 'Units Sold', 'Avg Price']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col, h in enumerate(headers, 1):
        cell = ws_rn.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, color="FFFFFF")

    # Data rows - realistic sales data for the North region
    data = [
        ['Laptop Pro 15', 'Electronics', 45200, 18, 2511.11],
        ['Wireless Mouse', 'Accessories', 8750, 125, 70.00],
        ['USB-C Hub', 'Accessories', 6300, 90, 70.00],
        ['Laptop Pro 15 Bundle', 'Electronics', 12000, 4, 3000.00],  # Row 5 → C5 = 12000
        ['Monitor 27"', 'Electronics', 23400, 12, 1950.00],
        ['Keyboard Mechanical', 'Accessories', 9800, 70, 140.00],
        ['Webcam HD', 'Peripherals', 4500, 50, 90.00],
        ['Docking Station', 'Accessories', 15600, 40, 390.00],
        ['Headset Pro', 'Peripherals', 11200, 80, 140.00],
        ['External SSD 1TB', 'Storage', 17500, 100, 175.00],
        ['Tablet Sleeve', 'Accessories', 3200, 160, 20.00],
        ['Power Bank 20K', 'Accessories', 5600, 80, 70.00],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws_rn.cell(row=r, column=c, value=val)

    # Format revenue column as currency
    for r in range(2, len(data) + 2):
        ws_rn.cell(row=r, column=3).number_format = '$#,##0.00'
        ws_rn.cell(row=r, column=5).number_format = '$#,##0.00'

    # Set column widths
    ws_rn.column_dimensions['A'].width = 24
    ws_rn.column_dimensions['B'].width = 14
    ws_rn.column_dimensions['C'].width = 14
    ws_rn.column_dimensions['D'].width = 12
    ws_rn.column_dimensions['E'].width = 12

    # --- Sheet 2: Dynamic ---
    ws_dyn = wb.create_sheet('Dynamic')

    # A1 = "Region", A2 = "_North", B1 = EMPTY (task requires agent to fill this)
    ws_dyn['A1'] = 'Region'
    ws_dyn['A2'] = '_North'
    # B1 intentionally left empty - the agent must create the INDIRECT formula here

    # Add some labels for context
    ws_dyn.cell(row=1, column=1).font = Font(bold=True)
    ws_dyn.cell(row=2, column=1).font = Font(bold=True)

    # Label column headers for clarity
    ws_dyn.cell(row=4, column=1, value='Parameter').font = Font(bold=True, underline='single')
    ws_dyn.cell(row=4, column=2, value='Value').font = Font(bold=True, underline='single')
    ws_dyn.cell(row=5, column=1, value='Sheet Name Part 1')
    ws_dyn.cell(row=5, column=2, value='Region')
    ws_dyn.cell(row=6, column=1, value='Sheet Name Part 2')
    ws_dyn.cell(row=6, column=2, value='_North')
    ws_dyn.cell(row=7, column=1, value='Target Cell')
    ws_dyn.cell(row=7, column=2, value='C5')

    ws_dyn.column_dimensions['A'].width = 22
    ws_dyn.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
