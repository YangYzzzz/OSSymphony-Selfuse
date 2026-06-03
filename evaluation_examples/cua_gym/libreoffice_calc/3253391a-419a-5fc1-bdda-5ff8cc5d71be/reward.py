"""
Reward Script: Calculate operating margin percentage for each product division and
               place a summary text in Sheet2 (Summary) that reads 'FY2023_Margin: X%'
Task ID: osworld_calc_gross_profit_sheet2_concat_011
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): Column D in DivisionFinancials sheet has operating margin formulas
                     for all 9 data rows (D2:D10), each computing (B-C)/B*100
  Component 2 (0.4): Summary sheet A1 contains a formula that concatenates 'FY2023_Margin:'
                     with the company-wide operating margin (TEXT(..., '0.00') & '%')
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_011'

# Sheet names as found in the workbook
MAIN_SHEET = 'DivisionFinancials'
SUMMARY_SHEET = 'Summary'

# Data rows: 2 through 10 (9 divisions)
DATA_ROWS = list(range(2, 11))


def normalize_formula(f):
    """Normalize formula for flexible comparison: uppercase, no spaces."""
    if f is None:
        return ''
    return str(f).upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Check both required sheets exist
    if MAIN_SHEET not in wb.sheetnames:
        print(f"PRECOND FAIL: Sheet '{MAIN_SHEET}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    if SUMMARY_SHEET not in wb.sheetnames:
        print(f"PRECOND FAIL: Sheet '{SUMMARY_SHEET}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_main = wb[MAIN_SHEET]
    ws_summary = wb[SUMMARY_SHEET]

    # -------------------------------------------------------------------------
    # Component 1: Column D has operating margin formulas for all 9 data rows (0.6 pts)
    # Each formula must compute operating margin: (Revenue - Costs) / Revenue * 100
    # Expected pattern: =(Bx-Cx)/Bx*100 or equivalent
    # This FAILS on initial (no D formulas) and PASSES on golden.
    # -------------------------------------------------------------------------
    try:
        formula_count = 0
        valid_formula_count = 0
        formula_details = []

        for row in DATA_ROWS:
            cell = ws_main.cell(row=row, column=4)  # column D
            val = cell.value
            if val is None:
                formula_details.append(f"D{row}: None")
                continue

            val_str = str(val).strip()
            if not val_str.startswith('='):
                formula_details.append(f"D{row}: not a formula: {repr(val_str)}")
                continue

            formula_count += 1
            norm = normalize_formula(val_str)

            # Accepted formula patterns (all equivalent):
            # =(Bx-Cx)/Bx*100
            # =(Bx-Cx)/Bx*100
            # Check that it references the correct row's B and C columns and computes margin
            # Pattern: starts with '=(B{row}-C{row})/B{row}*100' or similar

            b_col = f'B{row}'
            c_col = f'C{row}'

            # Check essential ingredients: B{row}, C{row}, division by B{row}, *100
            has_revenue = b_col.upper() in norm
            has_cost = c_col.upper() in norm
            has_multiply_100 = '*100' in norm
            has_division = '/' in norm

            if has_revenue and has_cost and has_multiply_100 and has_division:
                valid_formula_count += 1
                formula_details.append(f"D{row}: VALID ({val_str})")
            else:
                formula_details.append(f"D{row}: formula present but wrong pattern: {val_str}")

        print(f"\nComponent 1: Operating margin formulas in D2:D10")
        for detail in formula_details:
            print(f"  {detail}")
        print(f"  Valid formulas: {valid_formula_count}/{len(DATA_ROWS)}")

        if valid_formula_count == len(DATA_ROWS):
            # All 9 rows have correct operating margin formulas
            print(f"PASS: Component 1 — all {len(DATA_ROWS)} rows have valid operating margin formulas (0.6 pts)")
            total_score += 0.6
        elif valid_formula_count >= 5:
            # Partial credit: majority of formulas present
            partial = round(0.6 * valid_formula_count / len(DATA_ROWS), 2)
            print(f"PARTIAL: Component 1 — {valid_formula_count}/{len(DATA_ROWS)} valid formulas ({partial} pts)")
            total_score += partial
        elif formula_count > 0:
            # Some formulas present but not matching pattern
            print(f"PARTIAL: Component 1 — {formula_count} formula(s) in column D but not all valid (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 1 — no formulas in column D")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------------
    # Component 2: Summary sheet A1 contains FY2023_Margin concat formula (0.4 pts)
    # The formula must concatenate 'FY2023_Margin: ' with the company-wide margin
    # formatted to 2 decimal places, followed by '%'.
    # Expected: ="FY2023_Margin: "&TEXT(...margin formula...,"0.00")&"%"
    # This FAILS on initial (A1 is empty) and PASSES on golden.
    # -------------------------------------------------------------------------
    try:
        a1_val = ws_summary['A1'].value

        print(f"\nComponent 2: Summary A1 FY2023_Margin formula")
        print(f"  A1 value: {repr(a1_val)}")

        if a1_val is None or str(a1_val).strip() == '':
            print(f"FAIL: Component 2 — Summary A1 is empty")
        else:
            a1_str = str(a1_val).strip()
            norm_a1 = normalize_formula(a1_str)

            # The value could be:
            # 1. A formula string starting with '=' (e.g., ="FY2023_Margin: "&TEXT(...))
            # 2. A static computed string (e.g., "FY2023_Margin: 24.76%") if saved after calc
            # Check for formula form:
            is_formula = a1_str.startswith('=')

            # Check essential content
            has_fy2023 = 'FY2023_MARGIN' in norm_a1 or 'FY2023_MARGIN:' in norm_a1
            has_text_fn = 'TEXT(' in norm_a1 or 'TEXT(' in norm_a1.replace(' ', '')
            has_percent = '%' in a1_str
            has_sum_ref = 'SUM(' in norm_a1
            has_margin_calc = ('B2:B10' in norm_a1 or 'B2:B' in norm_a1) and \
                              ('C2:C10' in norm_a1 or 'C2:C' in norm_a1)

            # Also accept static value form: "FY2023_Margin: 24.76%"
            # In case the file was opened and saved by Calc (data_only cached result)
            static_pattern = re.match(r'FY2023_Margin:\s*\d+\.\d{2}%', a1_str, re.IGNORECASE)

            if is_formula and has_fy2023 and has_text_fn and has_percent and has_sum_ref and has_margin_calc:
                print(f"PASS: Component 2 — A1 has valid FY2023_Margin formula with TEXT and SUM (0.4 pts)")
                total_score += 0.4
            elif is_formula and has_fy2023 and has_percent and has_sum_ref:
                # Formula present with key elements but may be missing TEXT() formatting
                print(f"PARTIAL: Component 2 — A1 formula references FY2023_Margin and SUM but missing TEXT formatting (0.25 pts)")
                total_score += 0.25
            elif is_formula and has_fy2023 and has_percent:
                # Formula with FY2023 prefix and % but incomplete
                print(f"PARTIAL: Component 2 — A1 has FY2023_Margin formula but incomplete (0.15 pts)")
                total_score += 0.15
            elif static_pattern:
                # Static value that matches expected pattern (e.g., "FY2023_Margin: 24.76%")
                print(f"PASS: Component 2 — A1 contains static FY2023_Margin value: {a1_str} (0.4 pts)")
                total_score += 0.4
            elif 'FY2023' in a1_str.upper() or 'FY2023_MARGIN' in a1_str.upper():
                # Has some form of FY2023 reference
                print(f"PARTIAL: Component 2 — A1 mentions FY2023 but wrong format: {a1_str} (0.1 pts)")
                total_score += 0.1
            else:
                print(f"FAIL: Component 2 — A1 does not contain FY2023_Margin formula. Value: {repr(a1_str)}")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
