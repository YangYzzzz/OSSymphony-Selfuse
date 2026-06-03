"""
Initial Setup: Calculate operating margin percentage for product divisions
Task ID: osworld_calc_gross_profit_sheet2_concat_011
Domain: libreoffice_calc

Creates a workbook with Sheet1 containing product division financials
(Division, Revenue, Operating Costs) and Sheet2 as an empty summary sheet.
Column D in Sheet1 is intentionally left empty (no operating margin formulas).
Sheet2 A1 is also empty (no summary text yet).
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_gross_profit_sheet2_concat_011'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: DivisionFinancials ---
    ws1 = wb.active
    ws1.title = 'DivisionFinancials'

    # Headers: Division, Revenue, Operating Costs, Operating Margin %
    headers = ['Division', 'Revenue', 'Operating Costs', 'Operating Margin %']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # Realistic data: 9 product divisions (rows 2-10)
    # Column D (Operating Margin %) intentionally left empty — that's what the task adds
    data = [
        ['Consumer Electronics', 4850000, 3620000],
        ['Industrial Equipment',  3275000, 2480000],
        ['Software Solutions',    2940000, 1820000],
        ['Healthcare Devices',    2110000, 1650000],
        ['Automotive Parts',      3580000, 2890000],
        ['Renewable Energy',      1975000, 1430000],
        ['Smart Home Products',   2340000, 1760000],
        ['Aerospace Components',  4120000, 3280000],
        ['Agricultural Tech',     1685000, 1290000],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)
        # Column D (index 4) left empty intentionally

    # --- Sheet 2: Summary ---
    ws2 = wb.create_sheet('Summary')
    # Sheet2 exists but A1 is empty (agent will add the formula)
    ws2.cell(row=1, column=1, value=None)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the workbook in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
