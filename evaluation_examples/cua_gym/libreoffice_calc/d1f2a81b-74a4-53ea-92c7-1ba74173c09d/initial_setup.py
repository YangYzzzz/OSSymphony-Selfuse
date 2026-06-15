"""
Initial Setup: Export products.xlsx to CSV and analyze prices with VSCode Python script
Task ID: osworld_multi_apps_calc_vscode_007
Domain: libreoffice_calc + vscode (multi-app)
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_calc_vscode_007'
DESKTOP = f'{WORKDIR}/Desktop'
OUTPUT_XLSX = f'{DESKTOP}/products.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    # Ensure Desktop directory exists
    os.makedirs(DESKTOP, exist_ok=True)

    # Remove any leftover artifacts from a previous run
    for leftover in ['products.csv', 'stats.txt', 'analyze_products.py']:
        path = os.path.join(DESKTOP, leftover)
        if os.path.exists(path):
            os.remove(path)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Products'

    # --- Header row ---
    headers = ['Product', 'Category', 'Price', 'Quantity']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # --- Data rows (12 rows, 3 blank Price entries) ---
    # Price is None where blank is required
    data = [
        # Product,                  Category,       Price,  Quantity
        ['Wireless Keyboard',       'Electronics',  12.99,  150],
        ['Office Chair',            'Furniture',    34.50,  45],
        ['USB-C Hub',               'Electronics',  8.75,   200],
        ['Standing Desk Mat',       'Furniture',    None,   80],    # blank price
        ['Laptop Stand',            'Electronics',  22.00,  120],
        ['Blue Light Glasses',      'Accessories',  45.80,  60],
        ['Mechanical Pencils 12pk', 'Stationery',   None,   300],   # blank price
        ['Ergonomic Mouse',         'Electronics',  18.25,  175],
        ['Monitor Arm',             'Furniture',    67.90,  35],
        ['Desk Organizer',          'Stationery',   15.50,  250],
        ['Webcam HD 1080p',         'Electronics',  None,   90],    # blank price
        ['Notebook A5 Set',         'Stationery',   29.99,  400],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(OUTPUT_XLSX)
    print(f'Initial file created: {OUTPUT_XLSX}')

    # GUI-ready startup: open products.xlsx in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT_XLSX}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with products.xlsx on DISPLAY=:0')


create_initial()
