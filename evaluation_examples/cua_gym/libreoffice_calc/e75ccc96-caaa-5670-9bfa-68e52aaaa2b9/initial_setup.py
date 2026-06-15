"""
Initial Setup: Apply Currency cell style to monetary cells in Revenue Table
Task ID: calc_fmt_cell_style_currency_060
Domain: libreoffice_calc
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_cell_style_currency_060'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Revenue Table ---
    ws = wb.active
    ws.title = 'Revenue Table'

    # Headers (row 1)
    headers = ['Quarter', 'Revenue', 'Expenses', 'Net']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name='Calibri', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Quarterly financial data (rows 2-10) — 9 quarters of realistic business data
    # All values use General number format (NO currency formatting in initial)
    data = [
        ['Q1 2023', 245000, 189000, 56000],
        ['Q2 2023', 278500, 203400, 75100],
        ['Q3 2023', 312800, 241600, 71200],
        ['Q4 2023', 394200, 287300, 106900],
        ['Q1 2024', 267100, 198500, 68600],
        ['Q2 2024', 301400, 226800, 74600],
        ['Q3 2024', 348900, 259700, 89200],
        ['Q4 2024', 421600, 318200, 103400],
        ['Q1 2025', 289300, 214100, 75200],
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])  # Revenue — General format
        ws.cell(row=r, column=3, value=row_data[2])  # Expenses — General format
        ws.cell(row=r, column=4, value=row_data[3])  # Net — General format

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Revenue Table')
    print('Rows: 10 (1 header + 9 data rows)')
    print('B2:D10 format: General (no currency formatting)')


create_initial()
