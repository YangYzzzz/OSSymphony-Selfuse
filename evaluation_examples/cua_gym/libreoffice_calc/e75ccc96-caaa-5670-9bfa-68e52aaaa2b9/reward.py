"""
Reward Script: Apply 'Currency' cell style to B2:D10 in 'Revenue Table' sheet
Task ID: calc_fmt_cell_style_currency_060
Domain: libreoffice_calc
Scoring:
  Component 1 (0.40): All cells in B2:B10 have currency number format '$#,##0.00'
  Component 2 (0.30): All cells in C2:C10 have currency number format '$#,##0.00'
  Component 3 (0.30): All cells in D2:D10 have currency number format '$#,##0.00'
  Total: 1.0

The task requires applying the built-in 'Currency' cell style to all monetary value cells
in columns B, C, and D (rows 2-10). The Currency style uses format '$#,##0.00'.
Only cells B2:D10 should be modified; row 1 headers and column A should remain unchanged.
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_cell_style_currency_060'

# The currency number format used by LibreOffice's built-in 'Currency' style
CURRENCY_FORMAT = '$#,##0.00'


def verify_task(file_path):
    """
    Verify that cells B2:D10 in 'Revenue Table' sheet have the Currency number format applied.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: load the file
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet must exist
    if 'Revenue Table' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Revenue Table' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Revenue Table']

    # Component 1: B2:B10 — Revenue column has Currency format (0.40 points)
    # The initial file has 'General' for all cells; only golden should have '$#,##0.00'
    try:
        b_pass = 0
        b_total = 9
        b_failures = []
        for row in range(2, 11):
            cell = ws.cell(row=row, column=2)  # column B
            fmt = cell.number_format
            if fmt == CURRENCY_FORMAT:
                b_pass += 1
            else:
                b_failures.append(f"B{row}: found {repr(fmt)}")

        if b_pass == b_total:
            print(f"PASS: Component 1 — All {b_total} cells in B2:B10 have currency format '{CURRENCY_FORMAT}' (0.40 pts)")
            total_score += 0.40
        else:
            print(f"FAIL: Component 1 — Only {b_pass}/{b_total} cells in B2:B10 have currency format")
            if b_failures:
                print(f"  First failures: {b_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 1 (B2:B10 check) — {e}")

    # Component 2: C2:C10 — Expenses column has Currency format (0.30 points)
    try:
        c_pass = 0
        c_total = 9
        c_failures = []
        for row in range(2, 11):
            cell = ws.cell(row=row, column=3)  # column C
            fmt = cell.number_format
            if fmt == CURRENCY_FORMAT:
                c_pass += 1
            else:
                c_failures.append(f"C{row}: found {repr(fmt)}")

        if c_pass == c_total:
            print(f"PASS: Component 2 — All {c_total} cells in C2:C10 have currency format '{CURRENCY_FORMAT}' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 2 — Only {c_pass}/{c_total} cells in C2:C10 have currency format")
            if c_failures:
                print(f"  First failures: {c_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 (C2:C10 check) — {e}")

    # Component 3: D2:D10 — Net column has Currency format (0.30 points)
    try:
        d_pass = 0
        d_total = 9
        d_failures = []
        for row in range(2, 11):
            cell = ws.cell(row=row, column=4)  # column D
            fmt = cell.number_format
            if fmt == CURRENCY_FORMAT:
                d_pass += 1
            else:
                d_failures.append(f"D{row}: found {repr(fmt)}")

        if d_pass == d_total:
            print(f"PASS: Component 3 — All {d_total} cells in D2:D10 have currency format '{CURRENCY_FORMAT}' (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — Only {d_pass}/{d_total} cells in D2:D10 have currency format")
            if d_failures:
                print(f"  First failures: {d_failures[:3]}")
    except Exception as e:
        print(f"ERROR: Component 3 (D2:D10 check) — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
