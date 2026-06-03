"""
Reward Script: Weighted pipeline forecast column + bar chart by sales rep
Task ID: calc_sales_pipeline_winprob_002
Domain: libreoffice_calc

Scoring:
  Component 1: 'Weighted Value' header in Opportunities!H1            — 0.15 pts
  Component 2: =E*F formulas in Opportunities!H2:H81 (all 80 rows)   — 0.30 pts
  Component 3: 'Pipeline Chart' sheet exists                          — 0.15 pts
  Component 4: SUMIFS aggregation data (5 reps x rep+value columns)  — 0.15 pts
  Component 5: Bar/column chart present with correct title            — 0.25 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_sales_pipeline_winprob_002'


def get_title_text(title_obj):
    """Extract plain text from an openpyxl chart title object."""
    try:
        if title_obj is None:
            return None
        tx = title_obj.tx
        if tx and tx.rich:
            texts = []
            for p in tx.rich.p:
                for r in p.r:
                    if r.t:
                        texts.append(r.t)
            return ''.join(texts)
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Opportunities sheet must exist
    if 'Opportunities' not in wb.sheetnames:
        print("CRITICAL: 'Opportunities' sheet not found. Aborting.")
        print("REWARD: 0.0")
        return 0.0

    ws_opp = wb['Opportunities']

    # ------------------------------------------------------------------
    # Component 1: 'Weighted Value' header in Opportunities!H1 (0.15 pts)
    # This FAILS on initial (H1 is None) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        h1_value = ws_opp.cell(row=1, column=8).value
        if h1_value is not None and str(h1_value).strip() == 'Weighted Value':
            print(f"PASS: Component 1 — H1 header is 'Weighted Value' (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — Expected 'Weighted Value' in H1, found: {repr(h1_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: =E*F formulas in H2:H81 covering all 80 data rows (0.30 pts)
    # Formula pattern: =E<row>*F<row> or =F<row>*E<row> (order insensitive)
    # This FAILS on initial (H2:H81 are all None) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        formula_count = 0
        formula_pattern = re.compile(
            r'^=\s*(?:E(\d+)\s*\*\s*F\1|F(\d+)\s*\*\s*E\2)\s*$',
            re.IGNORECASE
        )
        for row in range(2, 82):  # rows 2–81
            cell_val = ws_opp.cell(row=row, column=8).value
            if cell_val is not None:
                cell_str = str(cell_val).strip()
                if formula_pattern.match(cell_str):
                    formula_count += 1

        if formula_count == 80:
            print(f"PASS: Component 2 — All 80 weighted value formulas (=E*F) found in H2:H81 (0.30 pts)")
            total_score += 0.30
        elif formula_count >= 40:
            # Partial: at least half the formulas are present
            partial = round(0.30 * (formula_count / 80), 2)
            print(f"PARTIAL: Component 2 — {formula_count}/80 weighted value formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {formula_count}/80 =E*F formulas found in H column")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: 'Pipeline Chart' sheet exists (0.15 pts)
    # This FAILS on initial (sheet does not exist) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        if 'Pipeline Chart' in wb.sheetnames:
            print(f"PASS: Component 3 — 'Pipeline Chart' sheet exists (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — 'Pipeline Chart' sheet not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: SUMIFS aggregation data on Pipeline Chart sheet (0.15 pts)
    # Must have 5 rep names and SUMIFS formulas referencing Opportunities!H
    # This FAILS on initial (sheet does not exist) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        if 'Pipeline Chart' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Pipeline Chart' sheet missing, cannot check aggregation")
        else:
            ws_pc = wb['Pipeline Chart']
            expected_reps = {'Sarah Chen', 'Mike Torres', 'Amy Liu', 'James Park', 'Rachel Green'}
            found_reps = set()
            sumifs_count = 0

            # Scan rows 2-10 for rep name data
            for row in range(2, 12):
                rep_name = ws_pc.cell(row=row, column=1).value
                formula_val = ws_pc.cell(row=row, column=2).value
                if rep_name and str(rep_name).strip() in expected_reps:
                    found_reps.add(str(rep_name).strip())
                if formula_val and isinstance(formula_val, str):
                    # Check for SUMIFS referencing Opportunities!H
                    if 'SUMIFS' in formula_val.upper() and 'Opportunities' in formula_val:
                        sumifs_count += 1

            reps_found = len(found_reps)
            if reps_found == 5 and sumifs_count >= 4:
                print(f"PASS: Component 4 — All 5 rep names found with SUMIFS aggregation (0.15 pts)")
                total_score += 0.15
            elif reps_found >= 3 and sumifs_count >= 3:
                partial = 0.10
                print(f"PARTIAL: Component 4 — {reps_found}/5 reps, {sumifs_count} SUMIFS formulas found ({partial} pts)")
                total_score += partial
            else:
                print(f"FAIL: Component 4 — Only {reps_found}/5 reps and {sumifs_count} SUMIFS formulas found")
                print(f"  Found reps: {found_reps}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ------------------------------------------------------------------
    # Component 5: Bar/column chart on 'Pipeline Chart' sheet with correct title (0.25 pts)
    # Chart must be a BarChart (col or bar type) titled 'Weighted Pipeline by Rep'
    # This FAILS on initial (sheet does not exist) and PASSES on golden
    # ------------------------------------------------------------------
    try:
        if 'Pipeline Chart' not in wb.sheetnames:
            print("FAIL: Component 5 — 'Pipeline Chart' sheet missing, cannot check chart")
        else:
            ws_pc = wb['Pipeline Chart']
            charts = ws_pc._charts

            if len(charts) == 0:
                print("FAIL: Component 5 — No chart found on 'Pipeline Chart' sheet")
            else:
                chart = charts[0]
                chart_type_name = type(chart).__name__

                # Check it is a BarChart (covers both col and bar subtypes)
                is_bar_chart = chart_type_name == 'BarChart'

                # Check title
                chart_title_text = get_title_text(chart.title)
                has_correct_title = (
                    chart_title_text is not None and
                    'Weighted Pipeline by Rep'.lower() in chart_title_text.lower()
                )

                if is_bar_chart and has_correct_title:
                    print(f"PASS: Component 5 — BarChart found with title '{chart_title_text}' (0.25 pts)")
                    total_score += 0.25
                elif is_bar_chart:
                    # Bar chart exists but title wrong — partial credit
                    print(f"PARTIAL: Component 5 — BarChart found but title '{chart_title_text}' "
                          f"does not match expected 'Weighted Pipeline by Rep' (0.15 pts)")
                    total_score += 0.15
                elif has_correct_title:
                    # Title correct but chart type wrong — partial credit
                    print(f"PARTIAL: Component 5 — Chart type '{chart_type_name}' found (expected BarChart) "
                          f"with correct title (0.15 pts)")
                    total_score += 0.15
                else:
                    print(f"FAIL: Component 5 — Chart type '{chart_type_name}', title '{chart_title_text}' "
                          f"— neither bar chart nor correct title")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
