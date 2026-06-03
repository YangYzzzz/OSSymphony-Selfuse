"""
Initial Setup: Sales pipeline opportunities spreadsheet
Task ID: calc_sales_pipeline_winprob_002
Domain: libreoffice_calc

Creates a workbook with an 'Opportunities' sheet containing 80 sales deals.
Column H is intentionally left empty — the agent must add the weighted value formula.
No 'Pipeline Chart' sheet exists — the agent must create it.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import random

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_pipeline_winprob_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Opportunities'

    # --- Headers (Row 1) ---
    headers = ['Opp ID', 'Rep Name', 'Account', 'Product', 'Deal Value', 'Win %', 'Quarter']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Column H header intentionally ABSENT (the task requires the agent to add 'Weighted Value')

    # --- Data rows (Rows 2-81: 80 opportunities) ---
    rep_names = ['Sarah Chen', 'Mike Torres', 'Amy Liu', 'James Park', 'Rachel Green']

    accounts = [
        'Acme Corp', 'TechNova Ltd', 'BlueStar Inc', 'Meridian Systems', 'GlobalEdge Co',
        'Apex Solutions', 'Frontier Analytics', 'Sterling Group', 'Pinnacle Networks', 'Vantage Partners',
        'Cascade Technologies', 'Summit Digital', 'NorthPoint Corp', 'Horizon Ventures', 'Clearwater LLC',
        'Redwood Consulting', 'Harbor Technologies', 'Lakewood Industries', 'Silverline Corp', 'Crestview Systems',
        'Ironclad Solutions', 'Brightstone Inc', 'Oceanview Analytics', 'Crossroads Media', 'Alpine Software',
        'Quantum Dynamics', 'Prism Analytics', 'Cobalt Enterprises', 'Ember Technologies', 'Zodiac Systems'
    ]

    products = [
        'CRM Pro Suite', 'Analytics Platform', 'Cloud Storage 500GB', 'Security Bundle',
        'Data Pipeline Tool', 'BI Dashboard', 'API Integration Pack', 'Support Tier Gold',
        'Enterprise License', 'DevOps Toolkit'
    ]

    quarters = ['Q1 2025', 'Q2 2025', 'Q3 2025', 'Q4 2025']

    # Fixed data for reproducibility
    random.seed(42)

    deal_values_pool = [
        15000, 22500, 31000, 45000, 52000, 67500, 78000, 89500, 95000, 102000,
        115000, 128000, 135000, 142000, 158000, 165000, 172000, 185000, 198000, 210000,
        220000, 235000, 248000, 255000, 262000, 270000, 280000, 18500, 27000, 38500,
        48000, 55000, 63000, 72000, 84000, 91000, 108000, 121000, 133000, 147000,
        160000, 175000, 189000, 205000, 215000, 228000, 242000, 258000, 265000, 275000
    ]

    win_pct_pool = [
        0.10, 0.15, 0.20, 0.25, 0.30, 0.35, 0.40, 0.45, 0.50,
        0.55, 0.60, 0.65, 0.70, 0.75, 0.80, 0.85, 0.90
    ]

    for i in range(80):
        row = i + 2
        opp_id = f'OPP-{1001 + i}'
        rep = rep_names[i % len(rep_names)]
        account = accounts[i % len(accounts)]
        product = products[i % len(products)]
        deal_val = deal_values_pool[i % len(deal_values_pool)]
        win_pct = win_pct_pool[i % len(win_pct_pool)]
        quarter = quarters[i % len(quarters)]

        ws.cell(row=row, column=1, value=opp_id)
        ws.cell(row=row, column=2, value=rep)
        ws.cell(row=row, column=3, value=account)
        ws.cell(row=row, column=4, value=product)
        ws.cell(row=row, column=5, value=deal_val)
        ws.cell(row=row, column=6, value=win_pct)
        ws.cell(row=row, column=7, value=quarter)
        # Column H (8) intentionally empty — agent must add =E*F formula

    # Format Deal Value column as currency
    for row in range(2, 82):
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6).number_format = '0%'

    # Column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Opportunities rows: 80 (rows 2-81)')
    print(f'  Column H: EMPTY (no header, no formulas)')

create_initial()
