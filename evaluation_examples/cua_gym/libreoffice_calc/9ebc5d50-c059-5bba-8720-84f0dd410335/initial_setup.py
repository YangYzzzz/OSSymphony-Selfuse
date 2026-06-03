"""
Initial Setup: NPS Survey Analysis Task
Task ID: calc_gen_analysis_035
Domain: libreoffice_calc

Creates a SurveyData sheet with 500 respondents.
Columns: Respondent ID, Segment, NPS Score, Classification (empty).
Classification column D is intentionally empty — agent must fill it.
No summary section, no chart, no conditional formatting.
"""

import os
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_analysis_035'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SurveyData'

    # --- Headers ---
    headers = ['Respondent ID', 'Segment', 'NPS Score', 'Classification']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')

    # --- Data ---
    segments = ['Enterprise', 'SMB', 'Consumer']
    # Distribute scores somewhat realistically:
    # Enterprise: higher scores (more promoters)
    # SMB: mixed
    # Consumer: lower scores

    score_distributions = {
        'Enterprise': [0,0,1,1,2,3,4,5,6,7,7,8,8,9,9,9,10,10,10,10],
        'SMB':        [0,1,2,3,4,5,5,6,6,7,7,8,8,8,9,9,9,10,10,10],
        'Consumer':   [0,1,1,2,3,3,4,4,5,5,6,6,7,7,8,8,9,9,10,10],
    }

    respondents = []
    # Generate exactly 500 respondents
    per_segment = {'Enterprise': 160, 'SMB': 170, 'Consumer': 170}
    rid = 1001
    for seg, count in per_segment.items():
        dist = score_distributions[seg]
        for _ in range(count):
            score = random.choice(dist)
            respondents.append((f'R{rid:04d}', seg, score))
            rid += 1

    # Shuffle respondents
    random.shuffle(respondents)
    # Re-assign IDs sequentially after shuffle
    respondents = [(f'R{i+1:04d}', seg, score) for i, (_, seg, score) in enumerate(respondents)]

    for i, (rid_val, seg, score) in enumerate(respondents):
        row = i + 2
        ws.cell(row=row, column=1, value=rid_val)
        ws.cell(row=row, column=2, value=seg)
        ws.cell(row=row, column=3, value=score)
        # Column D (Classification) intentionally left empty

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Column widths
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: SurveyData')
    print(f'  Rows: 501 (1 header + 500 data)')
    print(f'  Columns: Respondent ID, Segment, NPS Score, Classification (empty)')
    print(f'  Segments: Enterprise={per_segment["Enterprise"]}, SMB={per_segment["SMB"]}, Consumer={per_segment["Consumer"]}')

create_initial()
