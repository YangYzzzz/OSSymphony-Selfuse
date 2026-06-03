"""
Reward Script: NPS Analysis with Classification, Summary, Segment Breakdown, Chart, and Conditional Formatting
Task ID: calc_gen_analysis_035
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): D2:D501 filled with IF-formula classifying Promoter/Passive/Detractor
  Component 2 (0.25): NPS summary section (rows ~504-511) with COUNTIF and NPS formula
  Component 3 (0.25): Segment NPS breakdown table (rows ~514-518) with COUNTIFS and NPS per segment
  Component 4 (0.15): Bar chart present on SurveyData sheet
  Component 5 (0.10): Conditional formatting on segment NPS column (J516:J518)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_analysis_035'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: SurveyData sheet must exist
    if 'SurveyData' not in wb.sheetnames:
        print("CRITICAL: 'SurveyData' sheet not found.")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['SurveyData']

    # Component 1: Classification formulas in D2:D501 (0.25 points)
    # Initial state: D2:D501 all None (empty)
    # Golden state: IF formulas classifying 9-10 as Promoter, 7-8 as Passive, else Detractor
    try:
        formula_count = 0
        classified_count = 0
        for row in range(2, 502):  # rows 2 to 501
            val = ws.cell(row=row, column=4).value
            if val is not None:
                classified_count += 1
                if isinstance(val, str) and ('IF' in val.upper()):
                    formula_count += 1

        if formula_count >= 490:
            print(f"PASS: Component 1 — Classification formulas in D2:D501: {formula_count}/500 rows have IF formula (0.25 pts)")
            total_score += 0.25
        elif classified_count >= 490:
            print(f"PARTIAL: Component 1 — D column filled ({classified_count}/500 rows) but IF formula not detected. Giving 0.10 pts")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Only {classified_count}/500 rows in D column are filled (expected all 500 with IF formula)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: NPS Summary section at rows ~503-511 in column E/F (0.25 points)
    # Initial state: no summary section
    # Golden state: E504="NPS Summary", COUNTIF formulas for Promoter/Passive/Detractor counts,
    #               % Promoters, % Detractors, Overall NPS Score formula
    try:
        summary_row = None
        countif_count = 0

        # Scan rows 502-525 in column E for an NPS label
        for search_row in range(502, 525):
            cell_val = ws.cell(row=search_row, column=5).value
            if isinstance(cell_val, str) and 'NPS' in cell_val.upper():
                summary_row = search_row
                break

        if summary_row is not None:
            # Count COUNTIF formulas in the summary section
            for check_row in range(summary_row, summary_row + 12):
                for check_col in range(5, 8):  # columns E-G
                    v = ws.cell(row=check_row, column=check_col).value
                    if isinstance(v, str) and 'COUNTIF' in v.upper():
                        countif_count += 1

        if summary_row is not None and countif_count >= 2:
            print(f"PASS: Component 2 — NPS summary section found at row {summary_row} with {countif_count} COUNTIF formulas (0.25 pts)")
            total_score += 0.25
        elif summary_row is not None:
            print(f"PARTIAL: Component 2 — NPS summary section found at row {summary_row} but only {countif_count} COUNTIF formulas. Giving 0.12 pts")
            total_score += 0.12
        else:
            print(f"FAIL: Component 2 — No NPS summary section found in column E (rows 502-524)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Segment NPS breakdown table with COUNTIFS (0.25 points)
    # Initial state: no segment breakdown
    # Golden state: rows ~514-518 with Enterprise, SMB, Consumer labels, COUNTIFS formulas, NPS per segment
    try:
        segments_found = 0
        countifs_count = 0
        target_segments = {'Enterprise', 'SMB', 'Consumer'}

        # Scan rows 510-530 for segment labels and COUNTIFS formulas
        for check_row in range(510, 530):
            for check_col in range(5, 11):  # columns E-J
                cell_val = ws.cell(row=check_row, column=check_col).value
                if isinstance(cell_val, str):
                    if cell_val.strip() in target_segments:
                        segments_found += 1
                    if 'COUNTIFS' in cell_val.upper():
                        countifs_count += 1

        if segments_found >= 3 and countifs_count >= 3:
            print(f"PASS: Component 3 — Segment breakdown table: {segments_found} segment labels and {countifs_count} COUNTIFS formulas (0.25 pts)")
            total_score += 0.25
        elif segments_found >= 3:
            print(f"PARTIAL: Component 3 — Segment labels present ({segments_found}/3) but only {countifs_count} COUNTIFS formulas. Giving 0.12 pts")
            total_score += 0.12
        elif countifs_count >= 3:
            print(f"PARTIAL: Component 3 — COUNTIFS formulas present ({countifs_count}) but only {segments_found}/3 segment labels. Giving 0.12 pts")
            total_score += 0.12
        else:
            print(f"FAIL: Component 3 — Segment breakdown incomplete: {segments_found}/3 segments, {countifs_count} COUNTIFS formulas")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Bar chart present on SurveyData sheet (0.15 points)
    # Initial state: no charts (confirmed: len(ws._charts)==0 on initial)
    # Golden state: 1 BarChart with title "NPS Score by Segment"
    try:
        charts = ws._charts
        chart_count = len(charts)
        bar_chart_count = sum(1 for c in charts if 'Bar' in type(c).__name__)

        if bar_chart_count >= 1:
            print(f"PASS: Component 4 — Bar chart found on SurveyData sheet ({bar_chart_count} bar chart(s)) (0.15 pts)")
            total_score += 0.15
        elif chart_count >= 1:
            print(f"PARTIAL: Component 4 — Chart found but not a BarChart (type={type(charts[0]).__name__}). Giving 0.07 pts")
            total_score += 0.07
        else:
            print(f"FAIL: Component 4 — No chart found on SurveyData sheet (expected bar chart for NPS by segment)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Conditional formatting on NPS column for segments (0.10 points)
    # Initial state: no conditional formatting (confirmed: 0 CF rules on initial)
    # Golden state: conditional formatting on J516:J518 with green>=50%, yellow 0-49%, red<0
    try:
        cf_rules = ws.conditional_formatting
        cf_range_count = sum(1 for _ in cf_rules)
        cf_on_j_count = sum(
            1 for cfrange in cf_rules
            if 'J' in (str(cfrange.sqref) if hasattr(cfrange, 'sqref') else str(cfrange)).upper()
        )

        if cf_on_j_count >= 1:
            print(f"PASS: Component 5 — Conditional formatting on segment NPS (J column): {cf_range_count} rule set(s) (0.10 pts)")
            total_score += 0.10
        elif cf_range_count >= 1:
            print(f"PARTIAL: Component 5 — Conditional formatting present ({cf_range_count} rule sets) but not on J column. Giving 0.05 pts")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No conditional formatting found (expected CF on segment NPS scores)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
