"""
Reward Script: Generate randomized seating arrangement for final exam
Task ID: calc_edu_exam_seating_034
Domain: libreoffice_calc
Scoring:
  Component 1: RANK formula in column D (0.35 pts) — =RANK(Cx,$C$2:$C$151,1) for all 150 rows
  Component 2: Room assignment formula in column E (0.25 pts) — correct IF formula for all 150 rows
  Component 3: Seat number formula in column F (0.20 pts) — =MOD(Dx-1,50)+1 for all 150 rows
  Component 4: Data sorted by seating order ascending (0.20 pts) — C column has numeric values sorted ascending
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_exam_seating_034'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires:
    1. Column D: RANK formula =RANK(Cx,$C$2:$C$151,1) for rows 2-151
    2. Column E: Room assignment =IF(Dx<=50,"Room A",IF(Dx<=100,"Room B","Room C")) for rows 2-151
    3. Column F: Seat number =MOD(Dx-1,50)+1 for rows 2-151
    4. Data sorted by seating order (C column has numeric values, sorted ascending)
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Seating' sheet must exist
    if 'Seating' not in wb.sheetnames:
        print("FAIL: 'Seating' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Seating']

    # Precondition: check headers exist and basic sheet structure (rows 2-151)
    if ws.max_row < 151 or ws.max_column < 6:
        print(f"FAIL: Expected at least 151 rows and 6 columns, got rows={ws.max_row}, cols={ws.max_column}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Column D — RANK formula (0.35 points)
    # Expected pattern: =RANK(C{row},$C$2:$C$151,1) for rows 2-151
    try:
        rank_pattern = re.compile(r'^=RANK\(C(\d+),\$C\$2:\$C\$151,1\)$', re.IGNORECASE)
        correct_rank = 0
        total_rows = 150
        bad_rank_examples = []
        for row in range(2, 152):
            d_val = ws.cell(row=row, column=4).value
            if d_val is not None and isinstance(d_val, str):
                m = rank_pattern.match(d_val.strip())
                if m and m.group(1) == str(row):
                    correct_rank += 1
                else:
                    if len(bad_rank_examples) < 3:
                        bad_rank_examples.append((row, d_val))
            else:
                if len(bad_rank_examples) < 3:
                    bad_rank_examples.append((row, d_val))

        if correct_rank == total_rows:
            print(f"PASS: Component 1 — All 150 rows have correct RANK formula in column D (0.35 pts)")
            total_score += 0.35
        elif correct_rank >= 120:
            partial = round(0.35 * (correct_rank / total_rows), 4)
            print(f"PARTIAL: Component 1 — {correct_rank}/{total_rows} rows have correct RANK formula (+{partial} pts)")
            print(f"  Bad examples: {bad_rank_examples}")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {correct_rank}/{total_rows} rows have correct RANK formula")
            print(f"  Bad examples: {bad_rank_examples}")
    except Exception as e:
        print(f"ERROR: Component 1 (RANK formula check) — {e}")

    # Component 2: Column E — Room assignment formula (0.25 points)
    # Expected: =IF(D{row}<=50,"Room A",IF(D{row}<=100,"Room B","Room C"))
    try:
        e_pattern = re.compile(
            r'^=IF\(D(\d+)<=50,"Room A",IF\(D\1<=100,"Room B","Room C"\)\)$',
            re.IGNORECASE
        )
        correct_e = 0
        bad_e_examples = []
        for row in range(2, 152):
            e_val = ws.cell(row=row, column=5).value
            if e_val is not None and isinstance(e_val, str):
                m = e_pattern.match(e_val.strip())
                if m and m.group(1) == str(row):
                    correct_e += 1
                else:
                    if len(bad_e_examples) < 3:
                        bad_e_examples.append((row, e_val))
            else:
                if len(bad_e_examples) < 3:
                    bad_e_examples.append((row, e_val))

        if correct_e == total_rows:
            print(f"PASS: Component 2 — All 150 rows have correct Room assignment formula in column E (0.25 pts)")
            total_score += 0.25
        elif correct_e >= 120:
            partial = round(0.25 * (correct_e / total_rows), 4)
            print(f"PARTIAL: Component 2 — {correct_e}/{total_rows} rows have correct Room formula (+{partial} pts)")
            print(f"  Bad examples: {bad_e_examples}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {correct_e}/{total_rows} rows have correct Room formula")
            print(f"  Bad examples: {bad_e_examples}")
    except Exception as e:
        print(f"ERROR: Component 2 (Room formula check) — {e}")

    # Component 3: Column F — Seat number formula (0.20 points)
    # Expected: =MOD(D{row}-1,50)+1
    try:
        f_pattern = re.compile(r'^=MOD\(D(\d+)-1,50\)\+1$', re.IGNORECASE)
        correct_f = 0
        bad_f_examples = []
        for row in range(2, 152):
            f_val = ws.cell(row=row, column=6).value
            if f_val is not None and isinstance(f_val, str):
                m = f_pattern.match(f_val.strip())
                if m and m.group(1) == str(row):
                    correct_f += 1
                else:
                    if len(bad_f_examples) < 3:
                        bad_f_examples.append((row, f_val))
            else:
                if len(bad_f_examples) < 3:
                    bad_f_examples.append((row, f_val))

        if correct_f == total_rows:
            print(f"PASS: Component 3 — All 150 rows have correct Seat number formula in column F (0.20 pts)")
            total_score += 0.20
        elif correct_f >= 120:
            partial = round(0.20 * (correct_f / total_rows), 4)
            print(f"PARTIAL: Component 3 — {correct_f}/{total_rows} rows have correct Seat formula (+{partial} pts)")
            print(f"  Bad examples: {bad_f_examples}")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {correct_f}/{total_rows} rows have correct Seat formula")
            print(f"  Bad examples: {bad_f_examples}")
    except Exception as e:
        print(f"ERROR: Component 3 (Seat formula check) — {e}")

    # Component 4: Data sorted by seating order ascending (0.20 points)
    # In the golden file, C column has numeric values sorted ascending
    # (not =RAND() formulas anymore). The data was sorted by seating order.
    # We check that C column contains numeric values (not formulas) and they are sorted ascending.
    try:
        c_vals = []
        has_formulas = False
        for row in range(2, 152):
            c_val = ws.cell(row=row, column=3).value
            if isinstance(c_val, str) and c_val.strip().upper().startswith('='):
                has_formulas = True
                break
            elif isinstance(c_val, (int, float)):
                c_vals.append(c_val)
            else:
                # None or unexpected type
                c_vals.append(None)

        if has_formulas:
            print(f"FAIL: Component 4 — Column C still contains formulas (=RAND()), data was not sorted")
        elif len(c_vals) == 150 and all(v is not None for v in c_vals):
            # Check sorted ascending
            is_sorted = all(c_vals[i] <= c_vals[i + 1] for i in range(len(c_vals) - 1))
            if is_sorted:
                print(f"PASS: Component 4 — Column C has 150 numeric values sorted ascending (0.20 pts)")
                print(f"  Min: {c_vals[0]:.6f}, Max: {c_vals[-1]:.6f}")
                total_score += 0.20
            else:
                # Count how many adjacent pairs are out of order
                out_of_order = sum(1 for i in range(len(c_vals) - 1) if c_vals[i] > c_vals[i + 1])
                print(f"FAIL: Component 4 — Column C has {out_of_order} out-of-order adjacent pairs (not sorted)")
        else:
            none_count = sum(1 for v in c_vals if v is None)
            print(f"FAIL: Component 4 — Column C has {none_count} None values out of 150")
    except Exception as e:
        print(f"ERROR: Component 4 (Sort order check) — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
