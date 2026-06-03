"""
Initial Setup: Exam Seating Arrangement - Initial State
Task ID: calc_edu_exam_seating_034
Domain: libreoffice_calc

Creates a spreadsheet with 150 students with Student ID, Name, and RAND()
formulas in column C. Columns D (Seating Order), E (Room), F (Seat) are empty.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_exam_seating_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Seating'

    # Headers
    headers = ['Student ID', 'Name', 'Random Number', 'Seating Order', 'Room', 'Seat']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic student data - 150 students
    first_names = [
        'Emma', 'Liam', 'Olivia', 'Noah', 'Ava', 'Ethan', 'Sophia', 'Mason',
        'Isabella', 'William', 'Mia', 'James', 'Charlotte', 'Benjamin', 'Amelia',
        'Lucas', 'Harper', 'Henry', 'Evelyn', 'Alexander', 'Abigail', 'Sebastian',
        'Emily', 'Jack', 'Elizabeth', 'Michael', 'Mila', 'Owen', 'Ella', 'Samuel',
        'Avery', 'Daniel', 'Sofia', 'Matthew', 'Camila', 'Logan', 'Aria', 'Joseph',
        'Scarlett', 'David', 'Victoria', 'Carter', 'Madison', 'Owen', 'Luna',
        'Jayden', 'Grace', 'Dylan', 'Chloe', 'Grayson', 'Penelope', 'Lincoln',
        'Layla', 'Anthony', 'Riley', 'Isaac', 'Zoey', 'Eli', 'Nora', 'Connor',
        'Lily', 'Ezra', 'Eleanor', 'Aaron', 'Hannah', 'Charles', 'Lillian',
        'Thomas', 'Addison', 'Christopher', 'Aubrey', 'Jaxon', 'Ellie', 'Ryan',
        'Stella', 'Nathan', 'Natalie', 'Caleb', 'Zoe', 'Luke', 'Leah', 'Julian',
        'Hazel', 'Levi', 'Violet', 'Andrew', 'Aurora', 'Brayden', 'Savannah',
        'Brody', 'Audrey', 'Hunter', 'Brooklyn', 'Bentley', 'Bella', 'Adam',
        'Claire', 'Jason', 'Skylar', 'Xavier', 'Lucy', 'Josiah', 'Paisley',
        'Nolan', 'Everly', 'Lincoln', 'Anna', 'Elias', 'Caroline', 'Jordan',
        'Nova', 'Evan', 'Genesis', 'Carson', 'Emilia', 'Isaiah', 'Kennedy',
        'Maxwell', 'Samantha', 'Roman', 'Maya', 'Leo', 'Willow', 'Colton',
        'Kinsley', 'Miles', 'Naomi', 'Easton', 'Aaliyah', 'Theodore', 'Elena',
        'Dominic', 'Sarah', 'Jace', 'Ariana', 'Axel', 'Allison', 'Wesley',
        'Gabriella', 'Kayden', 'Alice', 'Camden', 'Hailey', 'Maverick', 'Eva',
        'Josue', 'Madelyn', 'Cooper', 'Gianna', 'Tristen', 'Peyton', 'Marcus'
    ]

    last_names = [
        'Smith', 'Johnson', 'Williams', 'Brown', 'Jones', 'Garcia', 'Miller',
        'Davis', 'Rodriguez', 'Martinez', 'Hernandez', 'Lopez', 'Gonzalez',
        'Wilson', 'Anderson', 'Thomas', 'Taylor', 'Moore', 'Jackson', 'Martin',
        'Lee', 'Perez', 'Thompson', 'White', 'Harris', 'Sanchez', 'Clark',
        'Ramirez', 'Lewis', 'Robinson', 'Walker', 'Young', 'Allen', 'King',
        'Wright', 'Scott', 'Torres', 'Nguyen', 'Hill', 'Flores', 'Green',
        'Adams', 'Nelson', 'Baker', 'Hall', 'Rivera', 'Campbell', 'Mitchell',
        'Carter', 'Roberts', 'Gomez', 'Phillips', 'Evans', 'Turner', 'Diaz',
        'Parker', 'Cruz', 'Edwards', 'Collins', 'Reyes', 'Stewart', 'Morris',
        'Morales', 'Murphy', 'Cook', 'Rogers', 'Gutierrez', 'Ortiz', 'Morgan',
        'Cooper', 'Peterson', 'Bailey', 'Reed', 'Kelly', 'Howard', 'Ramos',
        'Kim', 'Cox', 'Ward', 'Richardson', 'Watson', 'Brooks', 'Chavez',
        'Wood', 'James', 'Bennett', 'Gray', 'Mendoza', 'Ruiz', 'Hughes',
        'Price', 'Alvarez', 'Castillo', 'Sanders', 'Patel', 'Myers', 'Long',
        'Ross', 'Foster', 'Jimenez', 'Powell', 'Jenkins', 'Perry', 'Russell',
        'Sullivan', 'Bell', 'Coleman', 'Butler', 'Henderson', 'Barnes',
        'Gonzales', 'Fisher', 'Vasquez', 'Simmons', 'Romero', 'Jordan',
        'Patterson', 'Alexander', 'Hamilton', 'Graham', 'Reynolds', 'Griffin',
        'Wallace', 'Moreno', 'West', 'Cole', 'Hayes', 'Bryant', 'Herrera',
        'Gibson', 'Ellis', 'Tran', 'Medina', 'Aguilar', 'Stevens', 'Murray',
        'Ford', 'Castro', 'Marshall', 'Owens', 'Harrison', 'Fernandez', 'McDonald',
        'Woods', 'Washington', 'Kennedy', 'Wells', 'Vargas', 'Henry', 'Chen',
        'Freeman', 'Webb', 'Tucker', 'Guzman', 'Burns', 'Crawford', 'Olson'
    ]

    # Generate 150 students
    import random
    random.seed(42)  # For reproducible student list

    used_ids = set()
    for i in range(150):
        # Generate unique student ID
        while True:
            student_id = f'STU{random.randint(10000, 99999)}'
            if student_id not in used_ids:
                used_ids.add(student_id)
                break

        first = first_names[i % len(first_names)]
        last = last_names[i % len(last_names)]
        name = f'{first} {last}'

        row = i + 2
        ws.cell(row=row, column=1, value=student_id)
        ws.cell(row=row, column=2, value=name)
        # Column C: RAND() formula (already present, as specified in context)
        ws.cell(row=row, column=3, value=f'=RAND()')
        # Columns D, E, F: empty (to be filled by agent)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Rows: {ws.max_row - 1} students + 1 header')
    print(f'Columns D, E, F are empty (to be filled by agent)')

create_initial()
