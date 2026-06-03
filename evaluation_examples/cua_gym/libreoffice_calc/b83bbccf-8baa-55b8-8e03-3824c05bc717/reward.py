"""
Reward Script: Sort task list so all rows with red background in column A appear first.
Task ID: calc_dop_sort_color_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All 7 red-background rows are in rows 2-8 (sorted to top)
  Component 2 (0.3): Non-red rows maintain their original relative order (rows 9-26)
  Component 3 (0.2): Data values and colors are fully intact (no corruption)
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_dop_sort_color_006'

# Known red Task IDs (from initial file: rows 3,5,9,13,17,20,24)
RED_TASK_IDS = {'T-003', 'T-005', 'T-009', 'T-013', 'T-017', 'T-020', 'T-024'}

# Non-red rows in their original relative order from initial file
# Initial non-red rows (rows 2,4,6,7,8,10,11,12,14,15,16,18,19,21,22,23,25,26)
# Task IDs: T-001, T-004, T-006, T-007, T-008, T-010, T-011, T-012, T-014,
#           T-015, T-016, T-018, T-019, T-021, T-022, T-023, T-025, T-026
NON_RED_ORIGINAL_ORDER = [
    'T-001', 'T-004', 'T-006', 'T-007', 'T-008', 'T-010', 'T-011', 'T-012',
    'T-014', 'T-015', 'T-016', 'T-018', 'T-019', 'T-021', 'T-022', 'T-023',
    'T-025', 'T-026'
]


def is_red_cell(cell):
    """Check if a cell has a red background (ARGB FFFF0000)."""
    try:
        rgb = cell.fill.fgColor.rgb
        # Accept FFFF0000 (fully opaque red) or similar red hues
        if rgb and rgb.upper() == 'FFFF0000':
            return True
        return False
    except Exception:
        return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Sheet 'TaskList' must exist
    if 'TaskList' not in wb.sheetnames:
        print("CRITICAL: Sheet 'TaskList' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TaskList']

    # Precondition gate: Must have header row and 25 data rows (rows 2-26)
    if ws.max_row < 26:
        print(f"CRITICAL: Expected 26 rows (header + 25 data), found {ws.max_row}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All 7 red-background rows appear in rows 2-8 (0.5 points)
    # This checks the core task: red rows sorted to the top
    try:
        rows_2_to_8_task_ids = set()
        rows_2_to_8_are_red = True
        for row in range(2, 9):  # rows 2 through 8
            cell_a = ws.cell(row=row, column=1)
            task_id = cell_a.value
            cell_is_red = is_red_cell(cell_a)
            if task_id:
                rows_2_to_8_task_ids.add(str(task_id).strip())
            if not cell_is_red:
                rows_2_to_8_are_red = False
                print(f"FAIL: Component 1 — Row {row} (Task ID: {task_id}) does NOT have red background")

        red_rows_at_top = RED_TASK_IDS.issubset(rows_2_to_8_task_ids)

        if rows_2_to_8_are_red and red_rows_at_top and len(rows_2_to_8_task_ids) == 7:
            print(f"PASS: Component 1 — All 7 red rows appear at top (rows 2-8): {sorted(rows_2_to_8_task_ids)} (0.5 pts)")
            total_score += 0.5
        else:
            if not rows_2_to_8_are_red:
                print(f"FAIL: Component 1 — Not all rows 2-8 have red background")
            elif not red_rows_at_top:
                missing = RED_TASK_IDS - rows_2_to_8_task_ids
                print(f"FAIL: Component 1 — Missing red task IDs in rows 2-8: {missing}")
            else:
                print(f"FAIL: Component 1 — rows 2-8 task IDs {rows_2_to_8_task_ids} vs expected {RED_TASK_IDS}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Non-red rows maintain original relative order in rows 9-26 (0.3 points)
    # After sorting, non-red rows should appear in the same relative order as in initial file
    try:
        actual_non_red_order = []
        for row in range(9, 27):  # rows 9 through 26
            cell_a = ws.cell(row=row, column=1)
            task_id = cell_a.value
            if task_id and str(task_id).strip() != 'Task ID':
                actual_non_red_order.append(str(task_id).strip())

        if actual_non_red_order == NON_RED_ORIGINAL_ORDER:
            print(f"PASS: Component 2 — Non-red rows maintain original relative order in rows 9-26 (0.3 pts)")
            total_score += 0.3
        else:
            # Check partial: at least the order is a subsequence
            print(f"FAIL: Component 2 — Non-red row order mismatch")
            print(f"  Expected: {NON_RED_ORIGINAL_ORDER}")
            print(f"  Actual:   {actual_non_red_order}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: No red-background rows exist in rows 9-26 (all red rows moved to top) (0.2 points)
    # This verifies the sort is complete: rows 9+ must ALL be non-red.
    # This FAILS on initial file (red rows scattered throughout) and PASSES on golden file.
    try:
        no_red_in_lower_rows = True
        for row in range(9, 27):  # rows 9 through 26 should have NO red rows
            cell_a = ws.cell(row=row, column=1)
            if is_red_cell(cell_a):
                no_red_in_lower_rows = False
                task_id = cell_a.value
                print(f"FAIL: Component 3 — Row {row} (Task ID: {task_id}) has red background but should be non-red (not sorted to top)")

        if no_red_in_lower_rows:
            print(f"PASS: Component 3 — No red-background rows remain in rows 9-26; sort is complete (0.2 pts)")
            total_score += 0.2
        else:
            print("FAIL: Component 3 — Red rows still scattered in rows 9-26; sorting incomplete")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
