"""
Initial Setup: Sort task list by red background in column A
Task ID: calc_dop_sort_color_006
Domain: libreoffice_calc

Creates a TaskList sheet with 25 data rows (rows 2-26).
7 rows have red background in column A (high priority): rows 3, 5, 9, 13, 17, 20, 24
10 rows have yellow background (medium priority): remaining specified rows
8 rows have no fill (low priority)
Red-background rows are NOT sorted to the top (that is the task).
"""

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_sort_color_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Color fills
RED_FILL = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
NO_FILL = PatternFill(fill_type=None)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: TaskList ---
    ws = wb.active
    ws.title = "TaskList"

    # Headers in row 1
    headers = ['Task ID', 'Task Name', 'Assignee', 'Due Date', 'Status']
    bold_font = Font(bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = bold_font
        cell.alignment = Alignment(horizontal='center')

    # Column widths for readability
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    # Task data: 25 rows (rows 2-26)
    # Format: (task_id, task_name, assignee, due_date, status)
    # Row positions in final list (1-indexed from row 2):
    #   red rows (high priority): indices 1 (row2+1=row3), 3 (row5), 7 (row9), 11 (row13),
    #                              15 (row17), 18 (row20), 22 (row24)
    # i.e., data[0] is row2, data[1] is row3 (red), etc.

    task_data = [
        # row 2 - no fill (low priority)
        ('T-001', 'Update Employee Records',      'Alice Morgan',    '2025-03-10', 'Pending'),
        # row 3 - RED (high priority)
        ('T-003', 'Critical Server Issue',        'DevOps Team',     '2025-02-01', 'Open'),
        # row 4 - yellow (medium priority)
        ('T-004', 'Quarterly Budget Review',      'Finance Dept',    '2025-03-15', 'In Progress'),
        # row 5 - RED (high priority)
        ('T-005', 'Security Patch Deployment',    'IT Security',     '2025-01-28', 'Open'),
        # row 6 - yellow (medium priority)
        ('T-006', 'Marketing Campaign Setup',     'Emma Davis',      '2025-03-20', 'Pending'),
        # row 7 - no fill (low priority)
        ('T-007', 'Office Supply Reorder',        'Office Admin',    '2025-03-25', 'Pending'),
        # row 8 - yellow (medium priority)
        ('T-008', 'Software License Renewal',     'IT Department',   '2025-03-12', 'Pending'),
        # row 9 - RED (high priority)
        ('T-009', 'Data Breach Investigation',    'CISO Office',     '2025-02-05', 'Open'),
        # row 10 - no fill (low priority)
        ('T-010', 'Team Building Event Plan',     'HR Department',   '2025-04-01', 'Pending'),
        # row 11 - yellow (medium priority)
        ('T-011', 'Client Onboarding Process',    'Sales Team',      '2025-03-18', 'In Progress'),
        # row 12 - no fill (low priority)
        ('T-012', 'Website Content Update',       'Web Team',        '2025-03-22', 'Pending'),
        # row 13 - RED (high priority)
        ('T-013', 'Production Database Failure',  'DBA Team',        '2025-02-10', 'Open'),
        # row 14 - yellow (medium priority)
        ('T-014', 'Vendor Contract Negotiation',  'Procurement',     '2025-03-28', 'In Progress'),
        # row 15 - no fill (low priority)
        ('T-015', 'Internal Wiki Maintenance',    'Knowledge Mgmt',  '2025-04-05', 'Pending'),
        # row 16 - yellow (medium priority)
        ('T-016', 'Product Roadmap Planning',     'Product Mgmt',    '2025-03-30', 'In Progress'),
        # row 17 - RED (high priority)
        ('T-017', 'Network Outage Response',      'Network Ops',     '2025-02-15', 'Open'),
        # row 18 - no fill (low priority)
        ('T-018', 'Cafeteria Menu Revamp',        'Facilities',      '2025-04-10', 'Pending'),
        # row 19 - yellow (medium priority)
        ('T-019', 'Customer Feedback Analysis',   'CX Team',         '2025-03-24', 'In Progress'),
        # row 20 - RED (high priority)
        ('T-020', 'SSL Certificate Expiry',       'DevOps Team',     '2025-01-30', 'Open'),
        # row 21 - yellow (medium priority)
        ('T-021', 'Compliance Audit Prep',        'Legal Dept',      '2025-03-26', 'In Progress'),
        # row 22 - no fill (low priority)
        ('T-022', 'Parking Lot Resurfacing',      'Facilities',      '2025-04-15', 'Pending'),
        # row 23 - yellow (medium priority)
        ('T-023', 'API Integration Testing',      'QA Team',         '2025-03-19', 'In Progress'),
        # row 24 - RED (high priority)
        ('T-024', 'Firewall Configuration Error', 'Network Sec',     '2025-02-08', 'Open'),
        # row 25 - no fill (low priority)
        ('T-025', 'New Employee Orientation',     'HR Department',   '2025-04-08', 'Pending'),
        # row 26 - yellow (medium priority)
        ('T-026', 'Annual Performance Reviews',   'HR Department',   '2025-03-31', 'In Progress'),
    ]

    # Rows that should have RED fill (1-indexed in task_data, corresponds to spreadsheet rows 2-26)
    # task_data[0] = row2, task_data[1] = row3 (RED), etc.
    # Red rows: row3=idx1, row5=idx3, row9=idx7, row13=idx11, row17=idx15, row20=idx18, row24=idx22
    red_indices = {1, 3, 7, 11, 15, 18, 22}
    # Yellow rows: row4=idx2, row6=idx4, row8=idx6, row11=idx9, row14=idx12, row16=idx14,
    #              row19=idx17, row21=idx19, row23=idx21, row26=idx24
    yellow_indices = {2, 4, 6, 9, 12, 14, 17, 19, 21, 24}

    for idx, row_data in enumerate(task_data):
        row_num = idx + 2  # rows 2-26
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)

        # Apply background color to column A cell
        a_cell = ws.cell(row=row_num, column=1)
        if idx in red_indices:
            a_cell.fill = RED_FILL
        elif idx in yellow_indices:
            a_cell.fill = YELLOW_FILL
        # else: no fill (low priority)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: TaskList')
    print(f'  Rows: 1 header + 25 data rows')
    print(f'  Red rows (high priority) at data indices: {sorted(red_indices)} (rows: {[i+2 for i in sorted(red_indices)]})')
    print(f'  Yellow rows (medium priority) at data indices: {sorted(yellow_indices)}')
    print(f'  No-fill rows (low priority): remaining')

create_initial()
