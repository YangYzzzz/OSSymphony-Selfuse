"""
Reward Script: Multi-app email data processing task
Task ID: osworld_multi_apps_email_data_012
Domain: libreoffice_calc (multi-app)
Scoring:
  Component 1: 4 evaluation .txt files downloaded to /home/user/evaluations/ (0.20 pts)
  Component 2: rankings.csv created with correct evaluator ranking, top scorer is Dr. Sarah Rivera (0.30 pts)
  Component 3: xlsx Rankings sheet with bold headers, sorted descending by score, RANK formula in col D (0.35 pts)
  Component 4: Reply sent to top-ranked evaluator (Dr. Sarah Rivera) with correct congratulations text (0.15 pts)
  Total: 1.0
"""

import os

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_email_data_012'
EVALUATIONS_DIR = '/home/user/evaluations'
XLSX_PATH = f'/home/user/{TASK_ID}.xlsx'
RANKINGS_CSV = '/home/user/evaluations/rankings.csv'
THUNDERBIRD_DIR = '/home/user/.thunderbird'

# Expected evaluators sorted by total score descending
EXPECTED_TOP_EVALUATOR = 'Dr. Sarah Rivera'
EXPECTED_TOP_SCORE = 24
EXPECTED_EVALUATORS_BY_RANK = [
    ('Dr. Sarah Rivera', 24),
    ('Dr. Emily Watson', 23),
    ('Prof. Marcus Chen', 18),
    ("James O'Brien", 17),
]
EXPECTED_EVAL_FILES = [
    'evaluation_emily_watson.txt',
    'evaluation_james_obrien.txt',
    'evaluation_marcus_chen.txt',
    'evaluation_sarah_rivera.txt',
]


def verify_task():
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Component 1: 4 evaluation files downloaded to /home/user/evaluations/ (0.20 pts)
    # This check FAILS on initial_env (no evaluations dir) and PASSES on golden_env
    try:
        if not os.path.isdir(EVALUATIONS_DIR):
            print(f"FAIL: Component 1 — /home/user/evaluations/ directory does not exist")
        else:
            txt_files = [f for f in os.listdir(EVALUATIONS_DIR) if f.endswith('.txt')]
            if len(txt_files) >= 4:
                # Verify all expected files are present
                missing = [f for f in EXPECTED_EVAL_FILES if f not in txt_files]
                if not missing:
                    # Also verify each file has valid score content (5 criteria)
                    valid_files = 0
                    for fname in EXPECTED_EVAL_FILES:
                        fpath = os.path.join(EVALUATIONS_DIR, fname)
                        try:
                            with open(fpath, 'r') as fh:
                                content = fh.read()
                            lines = [l.strip() for l in content.splitlines() if ':' in l]
                            if len(lines) >= 5:
                                valid_files += 1
                        except Exception as fe:
                            print(f"  WARN: Could not read {fname}: {fe}")
                    if valid_files >= 4:
                        print(f"PASS: Component 1 — All 4 evaluation files present with valid score content (0.20 pts)")
                        total_score += 0.20
                    else:
                        print(f"FAIL: Component 1 — Only {valid_files}/4 files have valid 5-criteria content")
                else:
                    print(f"FAIL: Component 1 — Missing files: {missing}")
            else:
                print(f"FAIL: Component 1 — Only {len(txt_files)} .txt files found in evaluations/, expected 4")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: rankings.csv with correct content — top scorer Dr. Sarah Rivera with score 24 (0.30 pts)
    # This check FAILS on initial_env (no rankings.csv) and PASSES on golden_env
    try:
        if not os.path.exists(RANKINGS_CSV):
            print(f"FAIL: Component 2 — rankings.csv does not exist at {RANKINGS_CSV}")
        else:
            with open(RANKINGS_CSV, 'r') as fh:
                lines = [l.strip() for l in fh.readlines() if l.strip()]

            # Must have header + 4 data rows
            if len(lines) < 5:
                print(f"FAIL: Component 2 — rankings.csv has only {len(lines)} lines, expected 5 (header + 4 rows)")
            else:
                # Parse CSV rows (skip header)
                import csv
                import io
                reader = csv.DictReader(io.StringIO('\n'.join(lines)))
                rows = list(reader)

                if len(rows) < 4:
                    print(f"FAIL: Component 2 — rankings.csv has {len(rows)} data rows, expected 4")
                else:
                    # Check columns exist
                    has_evaluator = any(c.lower() in ('evaluator', 'name') for c in rows[0].keys())
                    has_score = any(c.lower() in ('totalscore', 'total_score', 'score', 'total') for c in rows[0].keys())

                    if not has_evaluator or not has_score:
                        print(f"FAIL: Component 2 — rankings.csv missing expected columns. Columns: {list(rows[0].keys())}")
                    else:
                        # Find score column name
                        score_col = None
                        evaluator_col = None
                        for c in rows[0].keys():
                            if c.lower() in ('totalscore', 'total_score', 'score', 'total'):
                                score_col = c
                            if c.lower() in ('evaluator', 'name'):
                                evaluator_col = c

                        # Check the first row (sorted desc) is Dr. Sarah Rivera with score 24
                        first_row = rows[0]
                        first_name = first_row.get(evaluator_col, '').strip()
                        try:
                            first_score = int(float(first_row.get(score_col, 0)))
                        except Exception:
                            first_score = -1

                        if first_name == EXPECTED_TOP_EVALUATOR and first_score == EXPECTED_TOP_SCORE:
                            print(f"PASS: Component 2 — rankings.csv has Dr. Sarah Rivera at top with score 24, sorted descending (0.30 pts)")
                            total_score += 0.30
                        else:
                            print(f"FAIL: Component 2 — Expected top entry: '{EXPECTED_TOP_EVALUATOR}' score 24, got: '{first_name}' score {first_score}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: xlsx Rankings sheet with bold headers, sorted descending by score, RANK formula in col D (0.35 pts)
    # This check FAILS on initial_env (no xlsx) and PASSES on golden_env
    try:
        import openpyxl
        if not os.path.exists(XLSX_PATH):
            print(f"FAIL: Component 3 — xlsx file does not exist at {XLSX_PATH}")
        else:
            wb = openpyxl.load_workbook(XLSX_PATH)

            # Check sheet named 'Rankings' exists
            if 'Rankings' not in wb.sheetnames:
                print(f"FAIL: Component 3 — No 'Rankings' sheet found in xlsx. Sheets: {wb.sheetnames}")
            else:
                ws = wb['Rankings']
                component3_score = 0.0
                component3_issues = []

                # Sub-check A: Bold headers in row 1 (A1, B1, C1)
                headers_bold = True
                for col in range(1, 4):
                    cell = ws.cell(row=1, column=col)
                    if not cell.font.bold:
                        headers_bold = False
                        component3_issues.append(f"Header cell {cell.coordinate} not bold")

                if headers_bold:
                    component3_score += 0.10
                    print(f"  PASS: Sub-check 3A — Header row is bold")
                else:
                    print(f"  FAIL: Sub-check 3A — {'; '.join(component3_issues)}")

                # Sub-check B: Data is sorted descending (row 2 has highest score)
                row2_b = ws.cell(row=2, column=2).value  # TotalScore of first data row
                row5_b = ws.cell(row=5, column=2).value  # TotalScore of last data row
                if row2_b is not None and row5_b is not None:
                    try:
                        if float(row2_b) > float(row5_b):
                            # Verify first row is Sarah Rivera with score 24
                            first_name = ws.cell(row=2, column=1).value
                            first_score = ws.cell(row=2, column=2).value
                            if str(first_name).strip() == EXPECTED_TOP_EVALUATOR and float(first_score) == EXPECTED_TOP_SCORE:
                                component3_score += 0.15
                                print(f"  PASS: Sub-check 3B — Sorted descending, Dr. Sarah Rivera (24) at top")
                            else:
                                print(f"  FAIL: Sub-check 3B — Sorted descending but top entry is '{first_name}' (score {first_score}), expected 'Dr. Sarah Rivera' (24)")
                        else:
                            print(f"  FAIL: Sub-check 3B — Data not sorted descending: row2={row2_b}, row5={row5_b}")
                    except Exception as se:
                        print(f"  ERROR: Sub-check 3B — {se}")
                else:
                    print(f"  FAIL: Sub-check 3B — Missing data in B2 or B5")

                # Sub-check C: RANK formula in column D (rows 2-5)
                rank_formulas_found = 0
                for row_num in range(2, 6):
                    cell = ws.cell(row=row_num, column=4)
                    val = cell.value
                    if isinstance(val, str) and 'RANK' in val.upper():
                        rank_formulas_found += 1

                if rank_formulas_found >= 4:
                    component3_score += 0.10
                    print(f"  PASS: Sub-check 3C — RANK formulas found in all 4 data rows in column D")
                elif rank_formulas_found >= 1:
                    component3_score += 0.05
                    print(f"  PARTIAL: Sub-check 3C — RANK formulas found in {rank_formulas_found}/4 rows in column D")
                else:
                    print(f"  FAIL: Sub-check 3C — No RANK formulas found in column D (rows 2-5)")

                if component3_score >= 0.35:
                    print(f"PASS: Component 3 — xlsx Rankings sheet complete ({component3_score:.2f} pts)")
                else:
                    print(f"PARTIAL: Component 3 — xlsx Rankings sheet incomplete ({component3_score:.2f}/0.35 pts)")
                total_score += component3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Reply email sent to Dr. Sarah Rivera with congratulations text (0.15 pts)
    # This check FAILS on initial_env (no Drafts or Sent with reply) and PASSES on golden_env
    try:
        CONGRATULATIONS_TEXT = 'Congratulations, you had the highest score!'
        REPLY_TO = 'sarah.rivera@research.edu'
        reply_found = False

        # Search in Thunderbird local mail folders for sent/drafts
        mail_base = os.path.join(THUNDERBIRD_DIR, 'b6x27ivi.default', 'Mail', 'Local Folders')
        for mail_file_name in ['Sent', 'Drafts', 'Outbox']:
            mail_file_path = os.path.join(mail_base, mail_file_name)
            if os.path.exists(mail_file_path):
                try:
                    with open(mail_file_path, 'r', encoding='utf-8', errors='replace') as mf:
                        content = mf.read()
                    if CONGRATULATIONS_TEXT in content and REPLY_TO in content:
                        reply_found = True
                        print(f"  PASS: Found congratulations reply to {REPLY_TO} in {mail_file_name}")
                        break
                    elif CONGRATULATIONS_TEXT in content:
                        # Check if the recipient matches loosely
                        if 'sarah' in content.lower() or 'rivera' in content.lower():
                            reply_found = True
                            print(f"  PASS: Found congratulations reply to Sarah Rivera in {mail_file_name}")
                            break
                except Exception as me:
                    print(f"  WARN: Could not read {mail_file_name}: {me}")

        if reply_found:
            print(f"PASS: Component 4 — Congratulations reply to Dr. Sarah Rivera found (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — No congratulations reply to Dr. Sarah Rivera found in Thunderbird Sent/Drafts")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score:.1f}")
    return final_score


verify_task()
