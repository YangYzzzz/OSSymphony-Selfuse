"""
Reward Script: Verify subtotal rows for each region in sales summary table
Task ID: calc_gsd_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Four subtotal rows exist with correct labels
  Component 2 (0.25): SUM formulas in Units column (D) for all subtotal rows
  Component 3 (0.25): SUM formulas in Revenue column (E) for all subtotal rows
  Component 4 (0.25): Bold formatting + light gray background on subtotal rows
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_009'

# Expected subtotal labels
EXPECTED_LABELS = ['North Total', 'South Total', 'East Total', 'West Total']


def find_subtotal_rows(ws):
    """Find rows whose column A value matches a subtotal label pattern."""
    found = {}  # label -> row number
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 1).value
        if val and str(val).strip() in EXPECTED_LABELS:
            found[str(val).strip()] = r
    return found


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'By Region' not in wb.sheetnames:
        print("FAIL: Sheet 'By Region' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['By Region']

    # Find subtotal rows dynamically
    subtotal_map = find_subtotal_rows(ws)

    # Component 1: Four subtotal rows exist with correct labels (0.25 points)
    try:
        found_count = 0
        for label in EXPECTED_LABELS:
            if label in subtotal_map:
                found_count += 1
                print(f"  Found '{label}' at row {subtotal_map[label]}")
            else:
                print(f"  MISSING: '{label}' not found in any row")

        if found_count == 4:
            print(f"PASS: Component 1 - All 4 subtotal labels found (0.25 pts)")
            total_score += 0.25
        elif found_count >= 2:
            partial = round(0.25 * found_count / 4, 3)
            print(f"PARTIAL: Component 1 - {found_count}/4 labels found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {found_count}/4 subtotal labels found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # If no subtotal rows found, no point checking further
    if not subtotal_map:
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: SUM formulas in Units column (D) for subtotal rows (0.25 points)
    try:
        units_pass = 0
        for label in EXPECTED_LABELS:
            if label not in subtotal_map:
                continue
            r = subtotal_map[label]
            val = ws.cell(r, 4).value  # Column D = Units
            if val and isinstance(val, str) and '=SUM(' in val.upper().replace(' ', ''):
                print(f"  Units SUM OK at row {r}: {val}")
                units_pass += 1
            else:
                print(f"  Units SUM MISSING at row {r}: found {val}")

        found_labels = len([l for l in EXPECTED_LABELS if l in subtotal_map])
        if found_labels > 0 and units_pass == found_labels:
            print(f"PASS: Component 2 - All {units_pass} Units SUM formulas correct (0.25 pts)")
            total_score += 0.25
        elif units_pass > 0:
            partial = round(0.25 * units_pass / max(found_labels, 1), 3)
            print(f"PARTIAL: Component 2 - {units_pass}/{found_labels} Units SUM formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No Units SUM formulas found in subtotal rows")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: SUM formulas in Revenue column (E) for subtotal rows (0.25 points)
    try:
        rev_pass = 0
        for label in EXPECTED_LABELS:
            if label not in subtotal_map:
                continue
            r = subtotal_map[label]
            val = ws.cell(r, 5).value  # Column E = Revenue
            if val and isinstance(val, str) and '=SUM(' in val.upper().replace(' ', ''):
                print(f"  Revenue SUM OK at row {r}: {val}")
                rev_pass += 1
            else:
                print(f"  Revenue SUM MISSING at row {r}: found {val}")

        found_labels = len([l for l in EXPECTED_LABELS if l in subtotal_map])
        if found_labels > 0 and rev_pass == found_labels:
            print(f"PASS: Component 3 - All {rev_pass} Revenue SUM formulas correct (0.25 pts)")
            total_score += 0.25
        elif rev_pass > 0:
            partial = round(0.25 * rev_pass / max(found_labels, 1), 3)
            print(f"PARTIAL: Component 3 - {rev_pass}/{found_labels} Revenue SUM formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No Revenue SUM formulas found in subtotal rows")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Bold formatting + light gray background on subtotal rows (0.25 points)
    try:
        style_pass = 0
        style_total = 0
        for label in EXPECTED_LABELS:
            if label not in subtotal_map:
                continue
            r = subtotal_map[label]
            style_total += 1
            bold_cells = sum(1 for c in range(1, 6) if ws.cell(r, c).font.bold)
            gray_cells = 0
            for c in range(1, 6):
                cell = ws.cell(r, c)
                try:
                    fg = cell.fill.fgColor.rgb
                    if fg in ('FFD9D9D9', 'FFC0C0C0', 'FFD3D3D3', 'FFBFBFBF'):
                        gray_cells += 1
                    elif fg and len(fg) == 8 and fg[:2] == 'FF':
                        r_val = int(fg[2:4], 16)
                        g_val = int(fg[4:6], 16)
                        b_val = int(fg[6:8], 16)
                        # Gray means R ~= G ~= B and value > 170 (light)
                        if abs(r_val - g_val) < 20 and abs(g_val - b_val) < 20 and r_val > 170:
                            gray_cells += 1
                except Exception:
                    pass
            row_bold = (bold_cells == 5)  # all 5 columns bold
            row_gray = (gray_cells == 5)  # all 5 columns gray bg

            if row_bold and row_gray:
                print(f"  Style OK for '{label}' at row {r}: bold + gray bg")
                style_pass += 1
            else:
                print(f"  Style ISSUE for '{label}' at row {r}: bold={row_bold}, gray={row_gray}")

        if style_total > 0 and style_pass == style_total:
            print(f"PASS: Component 4 - All {style_pass} subtotal rows have bold + gray bg (0.25 pts)")
            total_score += 0.25
        elif style_pass > 0:
            partial = round(0.25 * style_pass / max(style_total, 1), 3)
            print(f"PARTIAL: Component 4 - {style_pass}/{style_total} rows styled correctly ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 - No subtotal rows have correct bold + gray styling")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
