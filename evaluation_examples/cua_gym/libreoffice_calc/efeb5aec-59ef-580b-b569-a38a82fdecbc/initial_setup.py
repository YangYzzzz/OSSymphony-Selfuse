"""
Initial Setup: Add subtotal rows for each region in the sales summary table
Task ID: calc_gsd_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'By Region'

    # Headers in row 1
    headers = ['Rep', 'Region', 'Product', 'Units', 'Revenue']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # North region - rows 2-6
    north_data = [
        ['Sarah Chen',      'North', 'Widget A',    145, 7250.00],
        ['Marcus Johnson',  'North', 'Widget B',    210, 12600.00],
        ['Emily Park',      'North', 'Gadget Pro',   88, 6160.00],
        ['David Williams',  'North', 'Widget A',    176, 8800.00],
        ['Rachel Kim',      'North', 'Gadget Plus', 132, 9240.00],
    ]

    # South region - rows 7-11
    south_data = [
        ['James Carter',    'South', 'Widget B',    198, 11880.00],
        ['Maria Gonzalez',  'South', 'Gadget Pro',  156, 10920.00],
        ['Brian Foster',    'South', 'Widget A',    223, 11150.00],
        ['Olivia Thompson', 'South', 'Gadget Plus', 167, 11690.00],
        ['Kevin Lee',       'South', 'Widget B',     94, 5640.00],
    ]

    # East region - rows 12-16
    east_data = [
        ['Sophia Martinez', 'East', 'Gadget Pro',  185, 12950.00],
        ['Daniel Brown',    'East', 'Widget A',     241, 12050.00],
        ['Amanda White',    'East', 'Widget B',     119, 7140.00],
        ['Ryan Davis',      'East', 'Gadget Plus',  203, 14210.00],
        ['Jessica Taylor',  'East', 'Widget A',     157, 7850.00],
    ]

    # West region - rows 17-21
    west_data = [
        ['Michael Harris',  'West', 'Widget B',    172, 10320.00],
        ['Lisa Anderson',   'West', 'Gadget Pro',  138, 9660.00],
        ['Chris Robinson',  'West', 'Widget A',    265, 13250.00],
        ['Natalie Clark',   'West', 'Gadget Plus', 191, 13370.00],
        ['Tyler Wilson',    'West', 'Widget B',    114, 6840.00],
    ]

    all_data = north_data + south_data + east_data + west_data
    for r, row_data in enumerate(all_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
