"""
Initial Setup: Create a spreadsheet with employee data for header styling task
Task ID: calc_lf_071
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Report ---
    ws = wb.active
    ws.title = 'Report'

    # Headers (plain, no formatting)
    headers = ['Name', 'Department', 'Salary', 'Start Date', 'Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Employee data (realistic content, 10 rows)
    data = [
        ['Alice Wang', 'IT', 72000, '2023-01-15', 'Active'],
        ['Marcus Johnson', 'Marketing', 65000, '2022-06-01', 'Active'],
        ['Sarah Chen', 'Engineering', 91000, '2021-09-20', 'Active'],
        ['David Kim', 'Finance', 78000, '2023-03-10', 'On Leave'],
        ['Emily Rodriguez', 'HR', 62000, '2022-11-05', 'Active'],
        ['James O\'Brien', 'IT', 85000, '2020-07-22', 'Active'],
        ['Priya Patel', 'Engineering', 95000, '2021-01-30', 'Active'],
        ['Robert Taylor', 'Marketing', 58000, '2024-02-14', 'Probation'],
        ['Lisa Nakamura', 'Finance', 81000, '2022-04-18', 'Active'],
        ['Carlos Mendez', 'HR', 67000, '2023-08-07', 'Active'],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
