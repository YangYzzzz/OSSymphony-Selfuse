"""
Reward Script: Apply 'HeaderStyle' to header row A1:E1 with bold white text on dark blue background, centered.
Task ID: calc_lf_071
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3) - Bold font on all A1:E1
  Component 2 (0.25) - White font color (#FFFFFF) on all A1:E1
  Component 3 (0.25) - Dark blue background (#003366) on all A1:E1
  Component 4 (0.2) - Center horizontal alignment on all A1:E1
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_071'
HEADER_CELLS = ['A1', 'B1', 'C1', 'D1', 'E1']

# Expected color values (ARGB format as openpyxl stores them)
# Font color white: openpyxl stores 6-char input as 00FFFFFF (alpha 00), so accept both
EXPECTED_FONT_COLORS = {'00FFFFFF', 'FFFFFFFF'}
# Background dark blue #003366: should be stored as FF003366
EXPECTED_FILL_COLOR = 'FF003366'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check sheet exists
    if 'Report' not in wb.sheetnames:
        print("CRITICAL: 'Report' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Report']

    # Component 1: Bold font on all header cells A1:E1 (0.3 points)
    try:
        bold_count = 0
        for coord in HEADER_CELLS:
            cell = ws[coord]
            if cell.font.bold:
                bold_count += 1
            else:
                print(f"  DETAIL: {coord} bold={cell.font.bold}")
        if bold_count == 5:
            print(f"PASS: Component 1 — All 5 header cells are bold (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Only {bold_count}/5 header cells are bold")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: White font color on all header cells A1:E1 (0.25 points)
    try:
        white_count = 0
        for coord in HEADER_CELLS:
            cell = ws[coord]
            try:
                font_rgb = cell.font.color.rgb if cell.font.color else None
                if font_rgb and str(font_rgb) in EXPECTED_FONT_COLORS:
                    white_count += 1
                else:
                    print(f"  DETAIL: {coord} font_color={font_rgb}")
            except Exception:
                print(f"  DETAIL: {coord} font_color read error")
        if white_count == 5:
            print(f"PASS: Component 2 — All 5 header cells have white font color (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Only {white_count}/5 header cells have white font")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Dark blue background (#003366) on all header cells A1:E1 (0.25 points)
    try:
        fill_count = 0
        for coord in HEADER_CELLS:
            cell = ws[coord]
            try:
                fill_rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
                fill_type = cell.fill.patternType
                if fill_type == 'solid' and str(fill_rgb) == EXPECTED_FILL_COLOR:
                    fill_count += 1
                else:
                    print(f"  DETAIL: {coord} fill_fg={fill_rgb}, fill_type={fill_type}")
            except Exception:
                print(f"  DETAIL: {coord} fill read error")
        if fill_count == 5:
            print(f"PASS: Component 3 — All 5 header cells have dark blue background (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — Only {fill_count}/5 header cells have dark blue fill")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Center horizontal alignment on all header cells A1:E1 (0.2 points)
    try:
        center_count = 0
        for coord in HEADER_CELLS:
            cell = ws[coord]
            if cell.alignment.horizontal == 'center':
                center_count += 1
            else:
                print(f"  DETAIL: {coord} align_h={cell.alignment.horizontal}")
        if center_count == 5:
            print(f"PASS: Component 4 — All 5 header cells are center-aligned (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 4 — Only {center_count}/5 header cells are centered")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
