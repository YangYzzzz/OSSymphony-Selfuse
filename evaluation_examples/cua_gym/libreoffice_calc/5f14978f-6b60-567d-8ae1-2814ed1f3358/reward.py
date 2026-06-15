"""
Reward Script: Set up attendance tracking sheet with COUNTIF formulas, attendance rate, and conditional formatting.
Task ID: calc_edu_attendance_rate_002
Domain: libreoffice_calc
Scoring:
  Component 1: Column headers AL1:AO1 present (0.10 pts)
  Component 2: COUNTIF and rate formulas in AL2:AO26 for all 25 students (0.50 pts)
  Component 3: AO column formatted as percentage (0.10 pts)
  Component 4: Conditional formatting rule - yellow background for attendance < 90% (0.30 pts)
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_attendance_rate_002'

def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the Attendance sheet exists (precondition gate)
    if 'Attendance' not in wb.sheetnames:
        print("CRITICAL: 'Attendance' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Attendance']

    # Component 1: Column headers AL1='Absences', AM1='Tardies', AN1='Days Present', AO1='Attendance Rate' (0.10 pts)
    # These are NEW headers that don't exist in the initial file (which only has cols up to AK)
    try:
        al1 = ws.cell(row=1, column=38).value  # AL1
        am1 = ws.cell(row=1, column=39).value  # AM1
        an1 = ws.cell(row=1, column=40).value  # AN1
        ao1 = ws.cell(row=1, column=41).value  # AO1

        expected_headers = {
            'AL1': ('Absences', al1),
            'AM1': ('Tardies', am1),
            'AN1': ('Days Present', an1),
            'AO1': ('Attendance Rate', ao1),
        }

        headers_correct = all(
            expected == actual
            for _, (expected, actual) in expected_headers.items()
        )

        if headers_correct:
            print("PASS: Component 1 — All 4 column headers present (AL1='Absences', AM1='Tardies', AN1='Days Present', AO1='Attendance Rate') (0.10 pts)")
            total_score += 0.10
        else:
            missing = []
            for col_ref, (expected, actual) in expected_headers.items():
                if expected != actual:
                    missing.append(f"{col_ref}: expected {repr(expected)}, found {repr(actual)}")
            print(f"FAIL: Component 1 — Header mismatch: {'; '.join(missing)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: COUNTIF formulas in AL2:AN26 and rate formula in AO2:AO26 for all 25 students (0.50 pts)
    # Initial file has no values in these columns; golden has formulas for all 25 student rows.
    # Award partial credit proportionally: 0.02 pts per row with correct formulas (25 rows * 0.02 = 0.50)
    try:
        rows_correct = 0
        rows_partial = 0  # rows with at least some correct formulas
        formula_details = []

        for row in range(2, 27):  # rows 2-26 = 25 students
            expected_al = f'=COUNTIF(B{row}:AK{row},"A")'
            expected_am = f'=COUNTIF(B{row}:AK{row},"T")'
            expected_an = f'=COUNTIF(B{row}:AK{row},"P")'
            expected_ao = f'=AN{row}/40'

            al_val = ws.cell(row=row, column=38).value  # AL
            am_val = ws.cell(row=row, column=39).value  # AM
            an_val = ws.cell(row=row, column=40).value  # AN
            ao_val = ws.cell(row=row, column=41).value  # AO

            # Normalize formula comparison (case-insensitive, strip spaces)
            def norm(v):
                return str(v).upper().replace(' ', '') if v else ''

            al_ok = norm(al_val) == norm(expected_al)
            am_ok = norm(am_val) == norm(expected_am)
            an_ok = norm(an_val) == norm(expected_an)
            ao_ok = norm(ao_val) == norm(expected_ao)

            if al_ok and am_ok and an_ok and ao_ok:
                rows_correct += 1
            else:
                bad = []
                if not al_ok:
                    bad.append(f"AL{row}: expected {repr(expected_al)}, got {repr(al_val)}")
                if not am_ok:
                    bad.append(f"AM{row}: expected {repr(expected_am)}, got {repr(am_val)}")
                if not an_ok:
                    bad.append(f"AN{row}: expected {repr(expected_an)}, got {repr(an_val)}")
                if not ao_ok:
                    bad.append(f"AO{row}: expected {repr(expected_ao)}, got {repr(ao_val)}")
                formula_details.append(f"Row {row}: {'; '.join(bad)}")

        if rows_correct == 25:
            print(f"PASS: Component 2 — All 25 student rows have correct COUNTIF and attendance rate formulas (0.50 pts)")
            total_score += 0.50
        elif rows_correct > 0:
            partial = round(rows_correct * 0.02, 4)
            print(f"PARTIAL: Component 2 — {rows_correct}/25 student rows have correct formulas (+{partial} pts)")
            if formula_details:
                print(f"  First errors: {formula_details[0]}")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No student rows have correct formulas")
            if formula_details:
                print(f"  First error: {formula_details[0]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: AO column formatted as percentage (0.10 pts)
    # The initial file has no data in AO; golden should have '0.00%' or similar percentage format
    try:
        percentage_formats = 0
        for row in range(2, 27):
            ao_fmt = ws.cell(row=row, column=41).number_format  # AO
            # Accept any percentage-style format
            if ao_fmt and ('%' in ao_fmt):
                percentage_formats += 1

        if percentage_formats >= 25:
            print(f"PASS: Component 3 — AO column (Attendance Rate) formatted as percentage for all 25 rows (0.10 pts)")
            total_score += 0.10
        elif percentage_formats > 0:
            partial = round(percentage_formats * (0.10 / 25), 4)
            print(f"PARTIAL: Component 3 — {percentage_formats}/25 rows have percentage format in AO (+{partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — AO column not formatted as percentage (found: {repr(ws.cell(row=2, column=41).number_format)})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting rule — yellow background (#FFFF00 -> FFFFFF00 ARGB) for rows where AO < 0.90 (0.30 pts)
    # The initial file has no conditional formatting; golden has a formula-based rule
    # Award 0.15 for having a CF rule with the correct formula condition,
    # and 0.15 for having the yellow fill (FFFFFF00) in the DXF style
    try:
        cf_list = list(ws.conditional_formatting)
        if not cf_list:
            print("FAIL: Component 4 — No conditional formatting rules found on Attendance sheet")
        else:
            found_formula_rule = False
            found_yellow_fill = False
            rule_details = []

            for cf_obj in cf_list:
                for rule in cf_obj.rules:
                    rule_type = rule.type
                    formula = rule.formula if rule.formula else []

                    # Check for formula-based rule referencing AO column and < 0.9 threshold
                    formula_str = ' '.join(str(f) for f in formula).upper()
                    has_ao_ref = 'AO' in formula_str
                    has_threshold = '0.9' in formula_str or '0.90' in formula_str or '.9' in formula_str

                    if (rule_type == 'expression' or rule_type == 'formula') and has_ao_ref and has_threshold:
                        found_formula_rule = True

                    # Check for yellow fill in DXF
                    if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                        try:
                            fg_rgb = rule.dxf.fill.fgColor.rgb
                            # Accept FFFFFF00 (opaque yellow) or FFFF00 (6-char yellow)
                            if fg_rgb and ('FFFF00' in fg_rgb.upper()):
                                found_yellow_fill = True
                        except Exception:
                            pass

                    rule_details.append(f"type={rule_type}, formula={formula}")

            comp4_score = 0.0
            if found_formula_rule:
                comp4_score += 0.15
                print(f"PASS: Component 4a — CF rule with AO < 0.9 formula condition found (+0.15 pts)")
            else:
                print(f"FAIL: Component 4a — No CF formula rule referencing AO < 0.9 found. Rules: {rule_details}")

            if found_yellow_fill:
                comp4_score += 0.15
                print(f"PASS: Component 4b — CF rule has yellow fill (FFFFFF00) (+0.15 pts)")
            else:
                print(f"FAIL: Component 4b — CF rule does not have yellow fill. Rules: {rule_details}")

            total_score += comp4_score

            if comp4_score == 0.30:
                print(f"PASS: Component 4 — Complete conditional formatting: yellow background for attendance < 90% (0.30 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
