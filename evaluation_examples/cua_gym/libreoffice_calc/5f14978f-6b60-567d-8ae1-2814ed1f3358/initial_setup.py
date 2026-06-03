"""
Initial Setup: Attendance tracking sheet with 25 students and 36 school days (B:AK)
Task ID: calc_edu_attendance_rate_002
Domain: libreoffice_calc
"""

import openpyxl
import random
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_attendance_rate_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# Seed for reproducibility
random.seed(42)

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'

    # --- Row 1: Headers ---
    ws['A1'] = 'Student Name'

    # Generate school day date headers for columns B through AK (columns 2-37 = 36 columns)
    # These represent 36 school days recorded so far in a 40-day school year
    school_dates = []
    d = date(2024, 9, 3)
    while len(school_dates) < 36:
        if d.weekday() < 5:  # Mon-Fri only
            school_dates.append(d.strftime('%m/%d/%y'))
        d += timedelta(days=1)

    # Write dates to columns B(2) through AK(37)
    for i, dt in enumerate(school_dates):
        ws.cell(row=1, column=i + 2, value=dt)  # i+2: B=col2, ..., AK=col37

    # --- Rows 2-26: 25 Students ---
    student_names = [
        'Aaliyah Thompson', 'Benjamin Carter', 'Camila Rivera', 'Daniel Kim',
        'Evelyn Nguyen', 'Felix Okafor', 'Grace Liu', 'Henry Patel',
        'Isabella Martinez', 'Jasper Williams', 'Katherine Chen', 'Liam Johnson',
        'Maya Hernandez', 'Nathan Brooks', 'Olivia Davis', "Patrick O'Brien",
        'Quinn Robinson', 'Rachel Gonzalez', 'Samuel Jackson', 'Tiana Brown',
        'Uriel Santos', 'Victoria Lee', 'William Park', 'Xiomara Torres',
        'Yusuf Adeyemi'
    ]

    # Attendance pattern: some students below 90% (< 0.90 * 40 = 36 days, i.e., <36 present)
    # Actually the formula is AN/40 where AN = days present out of 36 data days
    # Below 90% means AN < 0.9 * 40 = 36, so less than 36 present (i.e., at least 1 absence/tardy)
    # Perfect attendance = 36/40 = 0.9 exactly (not < 0.9, so not flagged)
    # 1 absence = 35 present = 35/40 = 0.875 < 0.9 (flagged)
    # So we want some students with 1+ absences and some with 0 absences

    for row_idx, name in enumerate(student_names, 2):
        ws.cell(row=row_idx, column=1, value=name)

        # Vary attendance: row indices 3, 7, 15, 21 have poor attendance
        if row_idx in [3, 7, 15, 21]:
            # Poor attendance: 4-6 absences out of 36 days
            absence_count = random.randint(4, 6)
            tardy_count = random.randint(1, 2)
        elif row_idx in [5, 12, 18]:
            # Borderline: 1-2 absences
            absence_count = random.randint(1, 2)
            tardy_count = random.randint(0, 1)
        else:
            # Good attendance: 0 absences (present for all 36 days, rate = 0.9 exactly)
            absence_count = 0
            tardy_count = 0

        present_count = 36 - absence_count - tardy_count
        # Build attendance list
        attendance = ['P'] * present_count + ['A'] * absence_count + ['T'] * tardy_count
        random.shuffle(attendance)

        # Write to columns B(2) through AK(37)
        for col_offset, status in enumerate(attendance):
            ws.cell(row=row_idx, column=col_offset + 2, value=status)

    # Columns AL(38), AM(39), AN(40), AO(41) are intentionally left EMPTY
    # (These will be filled in by the agent with Absences, Tardies, Days Present, Attendance Rate)

    # Set column widths
    ws.column_dimensions['A'].width = 22
    for col_idx in range(2, 38):  # B through AK
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = 7

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Attendance')
    print(f'  Data columns: B through AK (36 school days)')
    print(f'  Students: rows 2-26 (25 students)')
    print(f'  Formula columns AL-AO: empty (to be added by agent)')

create_initial()
