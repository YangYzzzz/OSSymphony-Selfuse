"""
Initial Setup: Researcher venue analysis spreadsheet for DBLP data extraction task
Task ID: osworld_multi_apps_scholar_to_calc_014
Domain: libreoffice_calc

Creates venue_analysis.ods (as xlsx then converts) with:
- Section 1: Header row + 3 empty researcher rows (Name, Affiliation, Publications, Venue1, Venue2, Venue3)
- Section 2: 'Venue Overlap' heading with blank rows below
Both Chrome and LibreOffice Calc are launched GUI-ready.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_scholar_to_calc_014'
# Save as xlsx first, then convert to ods using LibreOffice headless
XLSX_PATH = f'{WORKDIR}/venue_analysis.xlsx'
ODS_PATH = f'{WORKDIR}/venue_analysis.ods'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def convert_to_ods(xlsx_path: str, ods_path: str):
    """Convert xlsx to ods using LibreOffice headless."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    result = subprocess.run(
        [
            'libreoffice', '--headless', '--convert-to', 'ods',
            '--outdir', WORKDIR, xlsx_path
        ],
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        env=env,
        timeout=30,
    )
    # Remove xlsx after conversion
    if os.path.exists(ods_path):
        try:
            os.remove(xlsx_path)
        except Exception:
            pass
    return os.path.exists(ods_path)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Researchers'

    # --- Section 1: Researcher data table ---
    # Row 1: Section heading
    ws['A1'] = 'Researcher Venue Analysis'
    ws['A1'].font = Font(bold=True, size=14)

    # Row 2: Column headers
    headers = ['Name', 'Affiliation', 'Publications', 'Venue 1', 'Venue 2', 'Venue 3']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Rows 3-5: Empty rows for 3 researchers (agent will fill these)
    for row_idx in range(3, 6):
        for col_idx in range(1, 7):
            ws.cell(row=row_idx, column=col_idx, value=None)

    # --- Section 2: Venue Overlap ---
    # Row 7: Section heading
    ws['A7'] = 'Venue Overlap'
    ws['A7'].font = Font(bold=True, size=12)
    ws['B7'] = '(Venues shared by 2+ researchers)'
    ws['B7'].font = Font(italic=True, color='FF555555')

    # Row 8: Sub-header
    ws['A8'] = 'Venue'
    ws['A8'].font = Font(bold=True)
    ws['B8'] = 'Researchers'
    ws['B8'].font = Font(bold=True)

    # Rows 9-14: Blank rows for agent to fill in shared venues
    for row_idx in range(9, 15):
        ws.cell(row=row_idx, column=1, value=None)
        ws.cell(row=row_idx, column=2, value=None)

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 22
    ws.column_dimensions['F'].width = 22

    wb.save(XLSX_PATH)
    print(f'Intermediate xlsx created: {XLSX_PATH}')

    # Convert to .ods
    success = convert_to_ods(XLSX_PATH, ODS_PATH)
    if success:
        print(f'Initial file created: {ODS_PATH}')
    else:
        # Fall back: keep xlsx if conversion failed
        print(f'ODS conversion result: {success}, keeping xlsx as fallback')
        # Rename xlsx to ods as a last resort so LibreOffice can open something
        if os.path.exists(XLSX_PATH) and not os.path.exists(ODS_PATH):
            import shutil
            shutil.copy(XLSX_PATH, ODS_PATH)
            print(f'Fallback: copied xlsx to {ODS_PATH}')

    # GUI-ready startup: open Chrome first, then LibreOffice Calc with the ods file
    # Open Chrome browser (agent will navigate to DBLP)
    launch_gui('google-chrome --new-window "https://dblp.org"', delay_sec=2.0)

    # Open LibreOffice Calc with the venue_analysis.ods file
    open_path = ODS_PATH if os.path.exists(ODS_PATH) else XLSX_PATH
    launch_gui(f'libreoffice --calc "{open_path}"', delay_sec=2.5)

    print('GUI_READY: launched Chrome and LibreOffice Calc with DISPLAY=:0')


create_initial()
