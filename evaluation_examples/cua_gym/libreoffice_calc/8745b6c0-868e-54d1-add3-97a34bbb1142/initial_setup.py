"""
Initial Setup: Build a project status dashboard
Task ID: calc_gcp_091
Domain: libreoffice_calc

Creates TaskTracker sheet with 150 rows of project task data.
Columns: TaskID, TaskName, Assignee, Status, DueDate, Priority
~18 tasks are Overdue. ~45% Completed.
NO pivot table, NO chart, NO conditional formatting (agent must create those).
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_091'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

# Deterministic seed for reproducibility
random.seed(42)

ASSIGNEES = [
    "Sarah Chen", "Marcus Johnson", "Priya Patel", "James O'Brien",
    "Yuki Tanaka", "Elena Rodriguez", "David Kim", "Fatima Al-Hassan",
    "Thomas Mueller", "Aisha Williams", "Ryan Cooper", "Mei Lin"
]

STATUSES = ["Not Started", "In Progress", "Under Review", "Completed", "Overdue"]

TASK_PREFIXES = [
    "Update", "Review", "Design", "Implement", "Test", "Deploy",
    "Configure", "Migrate", "Optimize", "Document", "Integrate",
    "Refactor", "Audit", "Create", "Validate"
]

TASK_SUBJECTS = [
    "user authentication module", "payment gateway", "dashboard UI",
    "API endpoints", "database schema", "search functionality",
    "notification system", "report generator", "file upload service",
    "cache layer", "logging framework", "CI/CD pipeline",
    "mobile responsive layout", "email templates", "analytics module",
    "access control policies", "backup procedures", "load balancer config",
    "SSL certificates", "container orchestration", "monitoring alerts",
    "data export feature", "onboarding flow", "billing integration",
    "inventory management", "customer portal", "vendor API connector",
    "compliance checks", "performance benchmarks", "security patches"
]

PRIORITIES = ["High", "Medium", "Low"]

def generate_task_name(idx):
    prefix = TASK_PREFIXES[idx % len(TASK_PREFIXES)]
    subject = TASK_SUBJECTS[idx % len(TASK_SUBJECTS)]
    return f"{prefix} {subject}"

def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TaskTracker"

    # Headers
    headers = ["TaskID", "TaskName", "Assignee", "Status", "DueDate", "Priority"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12

    # Distribution: ~67 Completed (~45%), ~25 In Progress, ~22 Not Started, ~18 Under Review, ~18 Overdue
    status_pool = (
        ["Completed"] * 67
        + ["In Progress"] * 25
        + ["Not Started"] * 22
        + ["Under Review"] * 18
        + ["Overdue"] * 18
    )
    random.shuffle(status_pool)

    today = date(2026, 4, 2)

    for i in range(150):
        row = i + 2
        task_id = f"T{i+1:03d}"
        task_name = generate_task_name(i)
        assignee = random.choice(ASSIGNEES)
        status = status_pool[i]
        priority = random.choice(PRIORITIES)

        # DueDate logic:
        # Completed: past dates
        # Overdue: past dates (these are overdue because they're past due and not completed)
        # In Progress, Under Review: mix of past and future
        # Not Started: future dates
        if status == "Completed":
            due = today - timedelta(days=random.randint(1, 90))
        elif status == "Overdue":
            due = today - timedelta(days=random.randint(3, 60))
        elif status == "Not Started":
            due = today + timedelta(days=random.randint(5, 90))
        elif status == "In Progress":
            due = today + timedelta(days=random.randint(-10, 45))
        else:  # Under Review
            due = today + timedelta(days=random.randint(-5, 30))

        ws.cell(row=row, column=1, value=task_id)
        ws.cell(row=row, column=2, value=task_name)
        ws.cell(row=row, column=3, value=assignee)
        ws.cell(row=row, column=4, value=status)
        cell_date = ws.cell(row=row, column=5, value=due)
        cell_date.number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=6, value=priority)

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")

create_initial()
