"""
Reward Script: Build a project status dashboard with pivot table, pie chart, and conditional formatting
Task ID: calc_gcp_091
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35) - Pivot table with Status counts in columns H-I
  Component 2 (0.35) - Pie chart created from pivot data
  Component 3 (0.30) - Conditional formatting highlighting Overdue rows in red
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_091'

# Expected status categories for the pivot table
EXPECTED_STATUSES = {'Completed', 'In Progress', 'Not Started', 'Overdue', 'Under Review'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that TaskTracker sheet exists
    if 'TaskTracker' not in wb.sheetnames:
        print("CRITICAL: 'TaskTracker' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['TaskTracker']

    # =========================================================================
    # Component 1: Pivot table with Status counts (0.35 points)
    # The pivot table should be a summary area with Status as row labels and
    # COUNT of tasks as values. We check for a region beyond the original F
    # column that contains all 5 status categories with numeric counts.
    # =========================================================================
    try:
        # Scan columns G onwards for a pivot-like table
        # Look for a header row containing "Status" and a count-related header
        pivot_found = False
        pivot_statuses_found = set()
        pivot_counts_valid = False
        status_col = None
        count_col = None

        # Search in columns 7-15 for pivot headers
        for col in range(7, 16):
            for row in range(1, 20):
                val = ws.cell(row=row, column=col).value
                if val and str(val).strip().lower() == 'status':
                    # Found status header, check if next column is count-like
                    next_val = ws.cell(row=row, column=col + 1).value
                    if next_val and str(next_val).strip().lower() in ('count', 'count of taskid', 'total', 'number', 'tasks'):
                        status_col = col
                        count_col = col + 1
                        pivot_found = True
                        break
            if pivot_found:
                break

        if pivot_found:
            # Read all status entries below the header
            header_row = None
            for row in range(1, 20):
                if ws.cell(row=row, column=status_col).value and str(ws.cell(row=row, column=status_col).value).strip().lower() == 'status':
                    header_row = row
                    break

            if header_row:
                for row in range(header_row + 1, header_row + 20):
                    status_val = ws.cell(row=row, column=status_col).value
                    count_val = ws.cell(row=row, column=count_col).value
                    if status_val is None:
                        break
                    status_str = str(status_val).strip()
                    if status_str.lower() == 'total':
                        continue  # skip total row
                    if status_str in EXPECTED_STATUSES:
                        pivot_statuses_found.add(status_str)

                # Check that counts are numeric and all 5 statuses present
                if len(pivot_statuses_found) >= 5:
                    # Verify at least one count is a positive integer
                    has_valid_count = False
                    for row in range(header_row + 1, header_row + 20):
                        count_val = ws.cell(row=row, column=count_col).value
                        if count_val is not None and isinstance(count_val, (int, float)) and count_val > 0:
                            has_valid_count = True
                            break
                    if has_valid_count:
                        pivot_counts_valid = True

        if pivot_found and len(pivot_statuses_found) >= 5 and pivot_counts_valid:
            print(f"PASS: Component 1 - Pivot table found with all 5 status categories and valid counts (0.35 pts)")
            total_score += 0.35
        elif pivot_found and len(pivot_statuses_found) >= 3:
            print(f"PARTIAL: Component 1 - Pivot table found but only {len(pivot_statuses_found)}/5 statuses: {pivot_statuses_found} (0.15 pts)")
            total_score += 0.15
        elif pivot_found:
            print(f"PARTIAL: Component 1 - Pivot table structure found but incomplete. Statuses: {pivot_statuses_found} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 - No pivot table found in columns G+ with Status/Count headers")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Pie chart from pivot data (0.35 points)
    # Should be a PieChart on the TaskTracker sheet
    # =========================================================================
    try:
        charts = ws._charts
        pie_chart_found = False
        pie_chart_title_ok = False

        for chart in charts:
            if isinstance(chart, openpyxl.chart.PieChart):
                pie_chart_found = True
                # Check title
                try:
                    if chart.title and chart.title.tx and chart.title.tx.rich:
                        for p in chart.title.tx.rich.paragraphs:
                            for r in p.r:
                                if r.t and 'status' in r.t.lower():
                                    pie_chart_title_ok = True
                except Exception:
                    pass
                break

        if not pie_chart_found:
            # Also check other sheets for a pie chart
            for sn in wb.sheetnames:
                if sn == 'TaskTracker':
                    continue
                other_ws = wb[sn]
                for chart in other_ws._charts:
                    if isinstance(chart, openpyxl.chart.PieChart):
                        pie_chart_found = True
                        try:
                            if chart.title and chart.title.tx and chart.title.tx.rich:
                                for p in chart.title.tx.rich.paragraphs:
                                    for r in p.r:
                                        if r.t and 'status' in r.t.lower():
                                            pie_chart_title_ok = True
                        except Exception:
                            pass
                        break
                if pie_chart_found:
                    break

        if pie_chart_found and pie_chart_title_ok:
            print(f"PASS: Component 2 - Pie chart found with status-related title (0.35 pts)")
            total_score += 0.35
        elif pie_chart_found:
            print(f"PARTIAL: Component 2 - Pie chart found but title does not mention status (0.25 pts)")
            total_score += 0.25
        else:
            # Check for any chart at all (might be a different chart type)
            any_chart = False
            for sn in wb.sheetnames:
                if len(wb[sn]._charts) > 0:
                    any_chart = True
                    break
            if any_chart:
                print(f"PARTIAL: Component 2 - Chart exists but is not a PieChart (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 2 - No charts found in the workbook")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Conditional formatting for Overdue rows (0.30 points)
    # Should apply a red background to rows where Status='Overdue'
    # =========================================================================
    try:
        cf_list = list(ws.conditional_formatting)
        overdue_cf_found = False
        red_fill_found = False

        for cf in cf_list:
            for rule in cf.rules:
                # Check if the rule references "Overdue" in its formula or condition
                is_overdue_rule = False

                if rule.type == 'expression' and rule.formula:
                    for f in rule.formula:
                        if 'overdue' in str(f).lower():
                            is_overdue_rule = True
                            break

                if rule.type == 'cellIs' and rule.formula:
                    for f in rule.formula:
                        if 'overdue' in str(f).lower():
                            is_overdue_rule = True
                            break

                if is_overdue_rule:
                    overdue_cf_found = True
                    # Check if the fill is red-ish
                    if rule.dxf and rule.dxf.fill:
                        try:
                            fg_rgb = rule.dxf.fill.fgColor.rgb
                            if fg_rgb:
                                rgb_str = str(fg_rgb).upper()
                                # Red = high R, low G, low B. ARGB format: AARRGGBB
                                # Check for common red variants
                                if 'FF0000' in rgb_str or 'FF4444' in rgb_str or 'CC0000' in rgb_str:
                                    red_fill_found = True
                                else:
                                    # Parse ARGB and check if red channel is dominant
                                    if len(rgb_str) == 8:
                                        r_val = int(rgb_str[2:4], 16)
                                        g_val = int(rgb_str[4:6], 16)
                                        b_val = int(rgb_str[6:8], 16)
                                        if r_val > 180 and g_val < 100 and b_val < 100:
                                            red_fill_found = True
                        except Exception:
                            pass
                    break

        if overdue_cf_found and red_fill_found:
            print(f"PASS: Component 3 - Conditional formatting for Overdue with red fill (0.30 pts)")
            total_score += 0.30
        elif overdue_cf_found:
            print(f"PARTIAL: Component 3 - Conditional formatting for Overdue found but fill is not clearly red (0.20 pts)")
            total_score += 0.20
        else:
            # Check if any conditional formatting exists at all
            if len(cf_list) > 0:
                print(f"PARTIAL: Component 3 - Conditional formatting exists but does not target Overdue rows (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 3 - No conditional formatting rules found")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
