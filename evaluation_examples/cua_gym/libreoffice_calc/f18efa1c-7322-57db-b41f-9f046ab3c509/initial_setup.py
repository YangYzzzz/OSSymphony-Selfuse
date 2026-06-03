"""
Initial Setup: Create spreadsheet with Revenue and Growth Rate data and a chart
Task ID: calc_chart_secondary_axis_052
Domain: libreoffice_calc

Creates an initial .xlsx file with:
- Sheet 'GrowthData' containing Year, Revenue ($M), Growth Rate % data
- A column chart with both Revenue and Growth Rate series on the SAME Y-axis
  (Growth Rate values are barely visible next to Revenue values - this is the problem to fix)
"""

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_secondary_axis_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: GrowthData ---
    ws = wb.active
    ws.title = 'GrowthData'

    # Headers
    ws['A1'] = 'Year'
    ws['B1'] = 'Revenue ($M)'
    ws['C1'] = 'Growth Rate %'

    # Style headers
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=12, color='FFFFFFFF')
    for col in ['A', 'B', 'C']:
        cell = ws[f'{col}1']
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Data rows (from task context)
    data = [
        (2020, 8.4, 0),
        (2021, 10.2, 21.4),
        (2022, 12.8, 25.5),
        (2023, 14.6, 14.1),
        (2024, 17.9, 22.6),
    ]

    for i, (year, revenue, growth) in enumerate(data, 2):
        ws[f'A{i}'] = year
        ws[f'B{i}'] = revenue
        ws[f'C{i}'] = growth

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16

    # --- Create a chart with BOTH series on the SAME Y-axis ---
    # (This is the initial state: Growth Rate values are barely visible
    #  because they share the same scale as Revenue values)
    chart = BarChart()
    chart.type = 'col'
    chart.title = 'Revenue and Growth Rate'
    chart.y_axis.title = 'Revenue ($M)'
    chart.x_axis.title = 'Year'
    chart.style = 10

    # Add both Revenue and Growth Rate as series (single chart, same Y-axis)
    # Both columns B and C are added to the same chart data range
    chart_data = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=6)
    chart.add_data(chart_data, titles_from_data=True)
    chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=6))

    # Place chart in the sheet
    ws.add_chart(chart, 'E2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Initial state: Both Revenue and Growth Rate share the same primary Y-axis')
    print('Problem: Growth Rate values (0-25) are barely visible next to Revenue values (8-18M)')


create_initial()
