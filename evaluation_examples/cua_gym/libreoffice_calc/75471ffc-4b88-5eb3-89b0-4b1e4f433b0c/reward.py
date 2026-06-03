"""
Reward Script: VLOOKUP + Pivot Table Combined Task
Task ID: osworld_calc_vlookup_pivot_combined_007
Domain: libreoffice_calc
Scoring:
  - Component 1: VLOOKUP formulas in E2:E21 of Sheet1 (0.40 pts)
  - Component 2: Sheet2 pivot table headers correct (0.20 pts)
  - Component 3: Sheet2 pivot data rows correct (0.30 pts)
  - Component 4: Sheet2 Grand Total row correct (0.10 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_007'


def _get_header_row_and_col_map(ws2):
    """Find the header row containing 'Category' and return (row_idx, col_map)."""
    for row_idx in range(1, 5):
        row_vals = [ws2.cell(row=row_idx, column=c).value for c in range(1, 8)]
        if row_vals[0] == 'Category':
            col_map = {}
            for col_idx, val in enumerate(row_vals, 1):
                if val is not None:
                    col_map[val] = col_idx
            return row_idx, col_map
    return None, {}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Fill Category column in Sheet1 via VLOOKUP (E2:E21), then create
    a two-dimensional pivot table in Sheet2 crossing Category by
    Sales Channel with total revenue and row/column Grand Totals.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: both Sheet1 and Sheet2 must exist
    if 'Sheet1' not in wb.sheetnames:
        print("FAIL: Sheet1 not found in workbook")
        print("REWARD: 0.0")
        return 0.0
    if 'Sheet2' not in wb.sheetnames:
        print("FAIL: Sheet2 not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws1 = wb['Sheet1']
    ws2 = wb['Sheet2']

    # -----------------------------------------------------------------------
    # Component 1: VLOOKUP formulas in Sheet1 column E (Category) (0.40 pts)
    # E2:E21 should all contain VLOOKUP formulas referencing $G$2:$H$9
    # In the initial file, E2:E21 are all None (empty).
    # -----------------------------------------------------------------------
    try:
        vlookup_count = 0
        non_formula_count = 0
        total_data_rows = 20  # rows 2-21

        for row_idx in range(2, 22):
            cell = ws1.cell(row=row_idx, column=5)  # column E
            val = cell.value
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(' ', '')
                # Must be a VLOOKUP formula referencing column B and the G:H lookup table
                if val_upper.startswith('=VLOOKUP(') and 'G' in val.upper() and 'H' in val.upper():
                    vlookup_count += 1
            elif val is not None:
                non_formula_count += 1

        if vlookup_count == total_data_rows:
            print(f"PASS: Component 1 — All {total_data_rows} VLOOKUP formulas found in E2:E21 (0.40 pts)")
            total_score += 0.40
        elif vlookup_count > 0:
            partial = round(0.40 * vlookup_count / total_data_rows, 4)
            print(f"PARTIAL: Component 1 — {vlookup_count}/{total_data_rows} VLOOKUP formulas in E column ({partial} pts)")
            if partial > 0:
                total_score += partial
        elif non_formula_count == total_data_rows:
            # Values entered but not via VLOOKUP formulas - give partial credit
            print(f"PARTIAL: Component 1 — {non_formula_count} direct category values in E column (no VLOOKUP formula, 0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Only {vlookup_count + non_formula_count} non-null values in E2:E21, expected {total_data_rows} VLOOKUP formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Sheet2 pivot table headers (0.20 pts)
    # Header row should have: Category, Online, Retail, Wholesale, Direct, Grand Total
    # -----------------------------------------------------------------------
    try:
        header_row, col_map = _get_header_row_and_col_map(ws2)

        if header_row is None:
            print("FAIL: Component 2 — No header row with 'Category' found in Sheet2")
        else:
            header_vals = [ws2.cell(row=header_row, column=c).value for c in range(1, 7)]
            required_headers = {'Category', 'Online', 'Retail', 'Wholesale', 'Direct', 'Grand Total'}
            found_headers = set(v for v in header_vals if v is not None)
            missing = required_headers - found_headers
            if not missing:
                print(f"PASS: Component 2 — Pivot headers correct: {header_vals} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 2 — Missing headers: {missing}, found: {header_vals}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Sheet2 pivot data rows (0.30 pts)
    # Expected rows (Category + per-channel revenue + row total):
    #   Electronics: Online=12450, Retail=25201, Wholesale=0, Direct=13231.25, Total=50882.25
    #   Clothing:    Online=11200, Retail=17740.25, Wholesale=11850, Direct=9875.25, Total=50665.5
    #   Home & Garden: Online=18750, Retail=7650.75, Wholesale=21000, Direct=10320, Total=57720.75
    #   Sports:      Online=44700, Retail=0, Wholesale=29000, Direct=0, Total=73700
    # -----------------------------------------------------------------------
    try:
        expected_data = {
            'Electronics':   {'Online': 12450.0, 'Retail': 25201.0,   'Wholesale': 0.0,     'Direct': 13231.25, 'Grand Total': 50882.25},
            'Clothing':      {'Online': 11200.0, 'Retail': 17740.25,  'Wholesale': 11850.0,  'Direct': 9875.25,  'Grand Total': 50665.5},
            'Home & Garden': {'Online': 18750.0, 'Retail': 7650.75,   'Wholesale': 21000.0,  'Direct': 10320.0,  'Grand Total': 57720.75},
            'Sports':        {'Online': 44700.0, 'Retail': 0.0,       'Wholesale': 29000.0,  'Direct': 0.0,      'Grand Total': 73700.0},
        }

        header_row_3, col_map_3 = _get_header_row_and_col_map(ws2)

        if header_row_3 is None or not col_map_3:
            print("FAIL: Component 3 — Cannot find header row in Sheet2")
        else:
            # Read data rows into dict (skip Grand Total row and title rows)
            actual_data = {}
            for row_idx in range(header_row_3 + 1, ws2.max_row + 1):
                cat_col = col_map_3.get('Category', 1)
                cat = ws2.cell(row=row_idx, column=cat_col).value
                if cat and cat != 'Grand Total':
                    row_dict = {}
                    for col_name, col_idx in col_map_3.items():
                        if col_name != 'Category':
                            row_dict[col_name] = ws2.cell(row=row_idx, column=col_idx).value
                    actual_data[cat] = row_dict

            # Score: each correct category row = 0.075 pts (4 categories * 0.075 = 0.30)
            pts_per_cat = round(0.30 / len(expected_data), 4)
            tolerance = 0.01
            cats_correct = 0

            for cat, exp_vals in expected_data.items():
                if cat not in actual_data:
                    print(f"FAIL: Component 3 — Category '{cat}' missing from pivot")
                    continue

                act_vals = actual_data[cat]
                mismatch_count = 0
                for channel, exp_val in exp_vals.items():
                    act_val = act_vals.get(channel)
                    if act_val is None:
                        act_val = 0.0
                    try:
                        act_float = float(act_val)
                    except (TypeError, ValueError):
                        act_float = 0.0
                    if abs(act_float - exp_val) > tolerance:
                        mismatch_count += 1
                        print(f"FAIL: Component 3 — {cat}/{channel}: expected {exp_val}, got {act_val}")

                if mismatch_count == 0:
                    cats_correct += 1
                    print(f"PASS: Component 3 — Category '{cat}' values correct (+{pts_per_cat:.4f} pts)")

            if cats_correct > 0:
                component3_score = round(cats_correct * pts_per_cat, 4)
                total_score += component3_score
                if cats_correct < len(expected_data):
                    print(f"PARTIAL: Component 3 — {cats_correct}/{len(expected_data)} categories correct ({component3_score} pts)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Sheet2 Grand Total row (0.10 pts)
    # Grand Total row: Online=87100, Retail=50592, Wholesale=61850, Direct=33426.5, Total=232968.5
    # -----------------------------------------------------------------------
    try:
        expected_grand_total = {
            'Online': 87100.0, 'Retail': 50592.0, 'Wholesale': 61850.0,
            'Direct': 33426.5, 'Grand Total': 232968.5
        }

        # Find Grand Total row
        grand_total_row = None
        for row_idx in range(1, ws2.max_row + 1):
            cat_val = ws2.cell(row=row_idx, column=1).value
            if cat_val == 'Grand Total':
                grand_total_row = row_idx
                break

        if grand_total_row is None:
            print("FAIL: Component 4 — 'Grand Total' row not found in Sheet2")
        else:
            _, col_map_4 = _get_header_row_and_col_map(ws2)
            tolerance = 0.01
            gt_mismatch_count = 0

            for channel, exp_val in expected_grand_total.items():
                col_idx = col_map_4.get(channel)
                if col_idx is None:
                    gt_mismatch_count += 1
                    print(f"FAIL: Component 4 — Column '{channel}' not found in header")
                    continue
                act_val = ws2.cell(row=grand_total_row, column=col_idx).value
                try:
                    act_float = float(act_val) if act_val is not None else 0.0
                except (TypeError, ValueError):
                    act_float = 0.0
                if abs(act_float - exp_val) > tolerance:
                    gt_mismatch_count += 1
                    print(f"FAIL: Component 4 — Grand Total/{channel}: expected {exp_val}, got {act_val}")

            if gt_mismatch_count == 0:
                print(f"PASS: Component 4 — Grand Total row correct (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 — Grand Total row has {gt_mismatch_count} mismatches")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
