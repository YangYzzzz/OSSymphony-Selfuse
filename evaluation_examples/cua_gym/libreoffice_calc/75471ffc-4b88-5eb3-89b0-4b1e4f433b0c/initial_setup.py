"""
Initial Setup: Product Performance Analysis - VLOOKUP + Pivot
Task ID: osworld_calc_vlookup_pivot_combined_007
Domain: libreoffice_calc

Creates a spreadsheet with:
  - Sheet1: Sales data (Sale ID, Product ID, Sales Channel, Revenue, Category)
    - Category column is EMPTY (agent must fill via VLOOKUP)
    - Reference table in columns G-H (Product ID -> Category)
  - Sheet2: Empty sheet (agent must create pivot table here)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_vlookup_pivot_combined_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ---- Sheet1: Sales Data ----
    ws1 = wb.active
    ws1.title = 'Sheet1'

    # Headers for sales table (columns A-E)
    headers = ['Sale ID', 'Product ID', 'Sales Channel', 'Revenue', 'Category']
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFFFF')
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Sales data - 20 realistic rows
    # Products: P001-P008 mapped to categories below
    # Sales Channels: Online, Retail, Wholesale, Direct
    sales_data = [
        ['S001', 'P001', 'Online',    12450.00, None],
        ['S002', 'P003', 'Retail',     8320.50, None],
        ['S003', 'P005', 'Wholesale', 15600.00, None],
        ['S004', 'P002', 'Direct',     9875.25, None],
        ['S005', 'P007', 'Online',    11200.00, None],
        ['S006', 'P004', 'Retail',     7650.75, None],
        ['S007', 'P006', 'Wholesale', 13400.00, None],
        ['S008', 'P001', 'Direct',     6980.50, None],
        ['S009', 'P008', 'Online',    18750.00, None],
        ['S010', 'P003', 'Retail',     9430.00, None],
        ['S011', 'P005', 'Online',    14200.00, None],
        ['S012', 'P002', 'Wholesale', 11850.00, None],
        ['S013', 'P007', 'Retail',     8640.25, None],
        ['S014', 'P004', 'Direct',    10320.00, None],
        ['S015', 'P006', 'Online',    16700.00, None],
        ['S016', 'P001', 'Retail',     7450.50, None],
        ['S017', 'P008', 'Wholesale', 21000.00, None],
        ['S018', 'P003', 'Direct',     6250.75, None],
        ['S019', 'P005', 'Online',    13800.00, None],
        ['S020', 'P002', 'Retail',     9100.00, None],
    ]

    for r, row_data in enumerate(sales_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = border
            if c == 4 and val is not None:
                cell.number_format = '#,##0.00'

    # Column widths for readability
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 15
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 16

    # Reference table in columns G-H: Product ID -> Category
    ref_headers = ['Product ID', 'Category']
    ref_header_fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    ref_header_font = Font(bold=True, color='FFFFFFFF')

    for col, h in enumerate(ref_headers, 7):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.fill = ref_header_fill
        cell.font = ref_header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ref_data = [
        ['P001', 'Electronics'],
        ['P002', 'Clothing'],
        ['P003', 'Electronics'],
        ['P004', 'Home & Garden'],
        ['P005', 'Sports'],
        ['P006', 'Sports'],
        ['P007', 'Clothing'],
        ['P008', 'Home & Garden'],
    ]

    for r, row_data in enumerate(ref_data, 2):
        for c, val in enumerate(row_data, 7):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = border

    ws1.column_dimensions['G'].width = 12
    ws1.column_dimensions['H'].width = 16

    # Freeze the header row
    ws1.freeze_panes = 'A2'

    # ---- Sheet2: Empty (agent builds pivot table here) ----
    ws2 = wb.create_sheet('Sheet2')
    ws2['A1'] = 'Pivot Table'
    ws2['A1'].font = Font(italic=True, color='FF888888')

    # Save
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
