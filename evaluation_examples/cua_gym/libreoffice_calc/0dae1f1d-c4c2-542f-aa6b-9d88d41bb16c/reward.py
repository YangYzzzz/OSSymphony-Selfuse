"""
Reward Script: Format numbers with custom 'units' suffix in LibreOffice Calc
Task ID: calc_lf_066
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): All three cells B2:B4 have number_format '#,##0" units"'
  Component 2 (0.3): Cells have correct numeric values AND the units format (compound)
  Component 3 (0.2): Values remain numeric type (not converted to strings) AND have units format
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_066'

# The expected custom number format
EXPECTED_FORMAT = '#,##0" units"'

# Expected cell values
EXPECTED_VALUES = {
    'B2': 1500,
    'B3': 340,
    'B4': 12800,
}


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify 'Inventory' sheet exists (precondition gate)
    if 'Inventory' not in wb.sheetnames:
        print("CRITICAL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Component 1: All three cells B2:B4 have the custom number format (0.5 points)
    # Award partial credit per cell: 0.5 / 3 per cell
    try:
        format_count = 0
        per_cell_pts = 0.5 / 3.0
        for coord in ['B2', 'B3', 'B4']:
            cell_fmt = ws[coord].number_format
            if cell_fmt == EXPECTED_FORMAT:
                format_count += 1
                total_score += per_cell_pts
                print(f"PASS: Component 1 — {coord} has format '{cell_fmt}' ({per_cell_pts:.4f} pts)")
            else:
                print(f"FAIL: Component 1 — {coord} expected format '{EXPECTED_FORMAT}', found '{cell_fmt}'")
        print(f"Component 1 summary: {format_count}/3 cells have correct format")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cells have correct numeric values AND the units format (0.3 points)
    # This is a compound check: value correctness anchored to the format change
    try:
        correct_count = 0
        per_cell_pts = 0.3 / 3.0
        for coord, expected_val in EXPECTED_VALUES.items():
            cell = ws[coord]
            has_format = cell.number_format == EXPECTED_FORMAT
            val = cell.value
            has_value = False
            if isinstance(val, (int, float)):
                has_value = abs(val - expected_val) < 0.01
            if has_format and has_value:
                correct_count += 1
                total_score += per_cell_pts
                print(f"PASS: Component 2 — {coord} value={val} with correct format ({per_cell_pts:.4f} pts)")
            else:
                print(f"FAIL: Component 2 — {coord} value={val!r} (expected {expected_val}), format='{cell.number_format}' (expected '{EXPECTED_FORMAT}')")
        print(f"Component 2 summary: {correct_count}/3 cells have correct value+format")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Values remain numeric type (not strings) AND have units format (0.2 points)
    # If someone typed "1500 units" as a string instead of applying format, this catches it
    try:
        numeric_with_format = 0
        per_cell_pts = 0.2 / 3.0
        for coord in ['B2', 'B3', 'B4']:
            cell = ws[coord]
            is_numeric = isinstance(cell.value, (int, float))
            has_format = cell.number_format == EXPECTED_FORMAT
            if is_numeric and has_format:
                numeric_with_format += 1
                total_score += per_cell_pts
                print(f"PASS: Component 3 — {coord} is numeric ({type(cell.value).__name__}) with units format ({per_cell_pts:.4f} pts)")
            else:
                print(f"FAIL: Component 3 — {coord} type={type(cell.value).__name__}, format='{cell.number_format}'")
        print(f"Component 3 summary: {numeric_with_format}/3 cells are numeric with format")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
