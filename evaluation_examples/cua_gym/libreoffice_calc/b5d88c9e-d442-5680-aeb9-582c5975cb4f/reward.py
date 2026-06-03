"""
Reward Script: Build a grade tracker with conditional formatting and a GPA calculator section.
Task ID: calc_gpm_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.15): Sheet "GradeTracker" with merged title A1:H1
  Component 2 (0.15): Header row 2 with correct column names, bold, navy fill
  Component 3 (0.15): Course data in rows 3-8 (names, credits, assignment avg, midterm)
  Component 4 (0.15): Course Grade formulas in F3:F8 (weighted avg =C*0.4+D*0.6)
  Component 5 (0.10): Letter grade formulas in G3:G8 (nested IF)
  Component 6 (0.10): GPA Points formulas in H3:H8 (letter-to-points mapping)
  Component 7 (0.10): GPA calculation in F10 with SUMPRODUCT formula, merged A10:E10
  Component 8 (0.10): Conditional formatting rules on G3:G8 and F10
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_033'


def persist_app_state(domain):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(0.8)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Sheet "GradeTracker" with merged title A1:H1 (0.15 points)
    try:
        if 'GradeTracker' in wb.sheetnames:
            ws = wb['GradeTracker']
            # Check title in A1
            title_val = ws['A1'].value
            # Check merge A1:H1
            merged_ranges = [str(r) for r in ws.merged_cells.ranges]
            has_title_merge = any('A1' in r and 'H1' in r for r in merged_ranges)
            if title_val and 'semester grade tracker' in str(title_val).lower() and has_title_merge:
                # Check bold and font size 14
                if ws['A1'].font.bold and ws['A1'].font.size and ws['A1'].font.size >= 14:
                    print(f"PASS: Component 1 — Sheet 'GradeTracker' with merged title A1:H1, bold 14pt (0.15 pts)")
                    total_score += 0.15
                elif title_val is not None:
                    print(f"PARTIAL: Component 1 — Title present and merged but font not bold 14pt (bold={ws['A1'].font.bold}, size={ws['A1'].font.size})")
                    total_score += 0.08
            else:
                print(f"FAIL: Component 1 — Expected 'Semester Grade Tracker' in merged A1:H1, found: value={title_val}, merge={has_title_merge}")
        else:
            print(f"FAIL: Component 1 — Sheet 'GradeTracker' not found. Sheets: {wb.sheetnames}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Get the worksheet for remaining checks
    if 'GradeTracker' not in wb.sheetnames:
        print("CRITICAL: No 'GradeTracker' sheet — cannot verify remaining components")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb['GradeTracker']

    # Component 2: Header row 2 with correct column names, bold, navy fill (0.15 points)
    try:
        expected_headers = ['Course', 'Credits', 'Assignment Avg', 'Midterm', 'Final',
                           'Course Grade', 'Letter', 'GPA Points']
        headers_found = 0
        for col_idx, expected in enumerate(expected_headers, 1):
            cell = ws.cell(row=2, column=col_idx)
            if cell.value and expected.lower() in str(cell.value).lower():
                headers_found += 1

        headers_bold = all(ws.cell(row=2, column=c).font.bold for c in range(1, 9))
        # Check navy fill (FF000080)
        try:
            fill_rgb = ws.cell(row=2, column=1).fill.fgColor.rgb
            has_navy_fill = fill_rgb and '000080' in str(fill_rgb)
        except Exception:
            has_navy_fill = False

        if headers_found >= 7 and headers_bold and has_navy_fill:
            print(f"PASS: Component 2 — All headers present, bold, navy fill ({headers_found}/8 match) (0.15 pts)")
            total_score += 0.15
        elif headers_found >= 6:
            print(f"PARTIAL: Component 2 — {headers_found}/8 headers, bold={headers_bold}, navy={has_navy_fill}")
            total_score += 0.08
        else:
            print(f"FAIL: Component 2 — Only {headers_found}/8 headers found, bold={headers_bold}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Course data in rows 3-8 (0.15 points)
    try:
        expected_courses = {
            3: ('Calculus II', 4, 88, 82),
            4: ('Organic Chem', 4, 75, 80),
            5: ('English Lit', 3, 92, 90),
            6: ('US History', 3, 85, 88),
            7: ('Physics I', 4, 78, 72),
            8: ('Comp Sci', 3, 95, 93),
        }
        courses_correct = 0
        for row, (name, credits, assign_avg, midterm) in expected_courses.items():
            a_val = ws.cell(row=row, column=1).value
            b_val = ws.cell(row=row, column=2).value
            c_val = ws.cell(row=row, column=3).value
            d_val = ws.cell(row=row, column=4).value
            if (a_val and name.lower() in str(a_val).lower() and
                b_val is not None and int(b_val) == credits and
                c_val is not None and int(c_val) == assign_avg and
                d_val is not None and int(d_val) == midterm):
                courses_correct += 1
            else:
                print(f"  Row {row}: A={a_val}, B={b_val}, C={c_val}, D={d_val} (expected {name},{credits},{assign_avg},{midterm})")

        if courses_correct == 6:
            print(f"PASS: Component 3 — All 6 courses with correct data (0.15 pts)")
            total_score += 0.15
        elif courses_correct >= 4:
            print(f"PARTIAL: Component 3 — {courses_correct}/6 courses correct")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — Only {courses_correct}/6 courses correct")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Course Grade formulas in F3:F8 (weighted avg =C*0.4+D*0.6) (0.15 points)
    try:
        formulas_correct = 0
        for row in range(3, 9):
            f_val = ws.cell(row=row, column=6).value
            if f_val and isinstance(f_val, str):
                f_upper = f_val.upper().replace(' ', '')
                # Check for weighted average pattern: C*0.4+D*0.6 or D*0.6+C*0.4
                has_c_weight = f'C{row}*0.4' in f_upper or f'0.4*C{row}' in f_upper
                has_d_weight = f'D{row}*0.6' in f_upper or f'0.6*D{row}' in f_upper
                if has_c_weight and has_d_weight:
                    formulas_correct += 1
                else:
                    print(f"  F{row}: formula={f_val} (does not match weighted avg pattern)")
            else:
                print(f"  F{row}: not a formula, value={f_val}")

        if formulas_correct == 6:
            print(f"PASS: Component 4 — All 6 Course Grade formulas correct (0.15 pts)")
            total_score += 0.15
        elif formulas_correct >= 4:
            print(f"PARTIAL: Component 4 — {formulas_correct}/6 formulas correct")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — Only {formulas_correct}/6 Course Grade formulas correct")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Letter grade formulas in G3:G8 (nested IF) (0.10 points)
    try:
        letter_formulas_ok = 0
        for row in range(3, 9):
            g_val = ws.cell(row=row, column=7).value
            if g_val and isinstance(g_val, str) and '=IF' in g_val.upper():
                # Check it references F column and contains letter grade strings
                g_upper = g_val.upper()
                if f'F{row}' in g_upper and '"A"' in g_val and '"B"' in g_val:
                    letter_formulas_ok += 1
                else:
                    print(f"  G{row}: IF formula present but doesn't reference F{row} or lacks grade letters")
            else:
                print(f"  G{row}: not an IF formula, value={g_val}")

        if letter_formulas_ok == 6:
            print(f"PASS: Component 5 — All 6 Letter grade formulas correct (0.10 pts)")
            total_score += 0.10
        elif letter_formulas_ok >= 4:
            print(f"PARTIAL: Component 5 — {letter_formulas_ok}/6 Letter formulas correct")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Only {letter_formulas_ok}/6 Letter formulas correct")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: GPA Points formulas in H3:H8 (letter-to-points mapping) (0.10 points)
    try:
        gpa_formulas_ok = 0
        for row in range(3, 9):
            h_val = ws.cell(row=row, column=8).value
            if h_val and isinstance(h_val, str) and '=IF' in h_val.upper():
                h_upper = h_val.upper()
                # Should reference G column and contain GPA point values like 4.0, 3.7, etc.
                if f'G{row}' in h_upper and '4' in h_val and '3.7' in h_val:
                    gpa_formulas_ok += 1
                else:
                    print(f"  H{row}: IF formula present but doesn't match GPA mapping pattern")
            else:
                print(f"  H{row}: not an IF formula, value={h_val}")

        if gpa_formulas_ok == 6:
            print(f"PASS: Component 6 — All 6 GPA Points formulas correct (0.10 pts)")
            total_score += 0.10
        elif gpa_formulas_ok >= 4:
            print(f"PARTIAL: Component 6 — {gpa_formulas_ok}/6 GPA Points formulas correct")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Only {gpa_formulas_ok}/6 GPA Points formulas correct")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: GPA calculation in F10 with SUMPRODUCT, merged A10:E10, bold label (0.10 points)
    try:
        # Check merged A10:E10
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_gpa_merge = any('A10' in r and 'E10' in r for r in merged_ranges)

        # Check label in A10
        a10_val = ws['A10'].value
        has_label = a10_val and 'semester gpa' in str(a10_val).lower()
        label_bold = ws['A10'].font.bold

        # Check SUMPRODUCT formula in F10
        f10_val = ws['F10'].value
        has_sumproduct = (f10_val and isinstance(f10_val, str) and
                         'SUMPRODUCT' in f10_val.upper() and
                         'SUM' in f10_val.upper())

        # Check number format (2 decimal places)
        f10_fmt = ws['F10'].number_format
        has_decimal_fmt = f10_fmt and ('0.00' in f10_fmt)

        if has_gpa_merge and has_label and has_sumproduct:
            detail = "merged A10:E10, label present, SUMPRODUCT formula"
            if label_bold:
                detail += ", label bold"
            if has_decimal_fmt:
                detail += ", 2-decimal format"
            print(f"PASS: Component 7 — {detail} (0.10 pts)")
            total_score += 0.10
        elif has_label and has_sumproduct:
            print(f"PARTIAL: Component 7 — label and SUMPRODUCT present but no A10:E10 merge")
            total_score += 0.05
        else:
            parts = []
            if not has_gpa_merge:
                parts.append("no A10:E10 merge")
            if not has_label:
                parts.append(f"label missing (A10={a10_val})")
            if not has_sumproduct:
                parts.append(f"no SUMPRODUCT formula (F10={f10_val})")
            print(f"FAIL: Component 7 — {', '.join(parts)}")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: Conditional formatting rules on G3:G8 and F10 (0.10 points)
    try:
        cf_ranges = []
        for cf in ws.conditional_formatting:
            cf_ranges.append(str(cf))

        has_g_cf = any('G3' in r for r in cf_ranges)
        has_f10_cf = any('F10' in r for r in cf_ranges)

        # Count total rules
        total_rules = 0
        for cf in ws.conditional_formatting:
            total_rules += len(cf.rules)

        if has_g_cf and has_f10_cf and total_rules >= 6:
            print(f"PASS: Component 8 — Conditional formatting on G3:G8 and F10, {total_rules} rules total (0.10 pts)")
            total_score += 0.10
        elif has_g_cf or has_f10_cf:
            partial = 0.05
            print(f"PARTIAL: Component 8 — CF on G={has_g_cf}, F10={has_f10_cf}, {total_rules} rules")
            total_score += partial
        else:
            print(f"FAIL: Component 8 — No conditional formatting found. Ranges: {cf_ranges}")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
