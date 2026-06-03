"""
Reward Script: AR Collections - Priority Tier and Formatting
Task ID: calc_fin_ar_collections_064
Domain: libreoffice_calc
Scoring:
  - Priority column with IF formula (G1 header + G2:G50 formulas): 0.25
  - Data sorted by Balance (D) descending: 0.20
  - Conditional formatting (High=red, Medium=yellow on A2:G50): 0.20
  - Freeze panes at C2 (first 2 cols + header row): 0.10
  - Row 1 bold headers: 0.05
  - D column currency formatted ($#,##0.00): 0.10
  - Comment on G2 saying 'Escalate to senior collector': 0.05
  - Summary row at row 52 with COUNTIF formulas: 0.05
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_ar_collections_064'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify the ARCollections sheet exists
    if 'ARCollections' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ARCollections' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ARCollections']

    # Component 1: Priority column header and IF formulas in G2:G50 (0.25 points)
    # G1 must be 'Priority', G2:G50 must contain IF formula referencing column D
    try:
        g1_val = ws.cell(1, 7).value
        has_priority_header = (g1_val is not None and str(g1_val).strip() == 'Priority')

        # Check G2:G50 have IF formulas
        formula_count = 0
        for row in range(2, 51):
            g_val = ws.cell(row, 7).value
            if g_val and isinstance(g_val, str) and g_val.upper().startswith('=IF(') and 'D' in g_val.upper():
                formula_count += 1

        if has_priority_header and formula_count >= 49:
            print(f"PASS: Component 1 — Priority column: G1='Priority', {formula_count}/49 IF formulas found (0.25 pts)")
            total_score += 0.25
        elif has_priority_header and formula_count >= 40:
            print(f"PARTIAL: Component 1 — G1='Priority' OK, only {formula_count}/49 IF formulas found (0.12 pts)")
            total_score += 0.12
        elif has_priority_header:
            print(f"PARTIAL: Component 1 — G1='Priority' OK but only {formula_count}/49 IF formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 1 — G1={repr(g1_val)}, IF formula count={formula_count}/49")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Data sorted by Balance (D) descending (0.20 points)
    # Check that D2:D50 are in descending order
    try:
        d_values = []
        for row in range(2, 51):
            val = ws.cell(row, 4).value
            if val is not None:
                try:
                    d_values.append(float(val))
                except (ValueError, TypeError):
                    pass

        if len(d_values) >= 10:
            is_sorted = all(d_values[i] >= d_values[i+1] for i in range(len(d_values)-1))
            if is_sorted:
                print(f"PASS: Component 2 — Data sorted by Balance descending ({len(d_values)} rows verified) (0.20 pts)")
                total_score += 0.20
            else:
                # Check if at least partially sorted (top half)
                half = len(d_values) // 2
                top_sorted = all(d_values[i] >= d_values[i+1] for i in range(half-1))
                if top_sorted:
                    print(f"PARTIAL: Component 2 — Top half sorted but not fully sorted (0.10 pts)")
                    total_score += 0.10
                else:
                    print(f"FAIL: Component 2 — Data not sorted descending by Balance. First 5: {d_values[:5]}")
        else:
            print(f"FAIL: Component 2 — Not enough Balance values found ({len(d_values)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on A2:G50 for High (red) and Medium (yellow) (0.20 points)
    # Expect formula-based rules: $G2="High" → fill red, $G2="Medium" → fill yellow
    try:
        cf_list = ws.conditional_formatting

        has_high_rule = False
        has_medium_rule = False
        high_color_ok = False
        medium_color_ok = False

        for cf in cf_list:
            cf_range = str(cf.sqref)
            for rule in cf.rules:
                formula_list = rule.formula if hasattr(rule, 'formula') and rule.formula else []
                for formula in formula_list:
                    formula_upper = formula.upper().replace(' ', '')
                    if '"HIGH"' in formula_upper or "'HIGH'" in formula_upper:
                        has_high_rule = True
                        # Check red fill
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                color_rgb = rule.dxf.fill.fgColor.rgb
                                # Red colors: FFFF9999 (light red), FFFF0000 (red), etc.
                                if color_rgb and ('FF' in color_rgb.upper()[:4] or
                                                  'FF9999' in color_rgb.upper() or
                                                  'FF0000' in color_rgb.upper() or
                                                  color_rgb.upper().startswith('FFFF')):
                                    high_color_ok = True
                            except Exception:
                                pass
                    elif '"MEDIUM"' in formula_upper or "'MEDIUM'" in formula_upper:
                        has_medium_rule = True
                        # Check yellow fill
                        if hasattr(rule, 'dxf') and rule.dxf and rule.dxf.fill:
                            try:
                                color_rgb = rule.dxf.fill.fgColor.rgb
                                # Yellow colors: FFFFFF99, FFFFFF00, etc.
                                if color_rgb and ('FFFF' in color_rgb.upper()[:4] or
                                                  'FFFF99' in color_rgb.upper() or
                                                  'FFFF00' in color_rgb.upper()):
                                    medium_color_ok = True
                            except Exception:
                                pass

        if has_high_rule and has_medium_rule and high_color_ok and medium_color_ok:
            print("PASS: Component 3 — CF rules for High (red) and Medium (yellow) on A2:G50 (0.20 pts)")
            total_score += 0.20
        elif has_high_rule and has_medium_rule:
            print(f"PARTIAL: Component 3 — Both CF rules found but color check: high_color={high_color_ok}, medium_color={medium_color_ok} (0.10 pts)")
            total_score += 0.10
        elif has_high_rule or has_medium_rule:
            print(f"PARTIAL: Component 3 — Only one CF rule found: high={has_high_rule}, medium={has_medium_rule} (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules found for High/Medium priorities")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Freeze panes at C2 (first 2 columns and header row frozen) (0.10 points)
    try:
        fp = ws.freeze_panes
        if fp == 'C2':
            print(f"PASS: Component 4 — Freeze panes at C2 (0.10 pts)")
            total_score += 0.10
        elif fp in ('C1', 'B2', 'B1', 'A2'):
            # Partial credit for partial freeze
            print(f"PARTIAL: Component 4 — Freeze panes at {fp} (not C2) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — Freeze panes={repr(fp)}, expected 'C2'")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Row 1 headers are bold (0.05 points)
    # Task requires "Row 1 bold" — headers should be bolded
    try:
        bold_count = 0
        for col in range(1, 8):
            if ws.cell(1, col).font.bold:
                bold_count += 1
        if bold_count >= 6:
            print(f"PASS: Component 5 — Row 1 headers bold ({bold_count}/7 cells) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — Only {bold_count}/7 header cells are bold")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: D column (Balance) currency formatted with $#,##0.00 (0.10 points)
    try:
        currency_count = 0
        for row in range(2, 51):
            fmt = ws.cell(row, 4).number_format
            # Accept $#,##0.00 or similar currency formats
            if fmt and ('$' in fmt or '0.00' in fmt) and fmt != 'General':
                currency_count += 1

        if currency_count >= 49:
            print(f"PASS: Component 6 — D column currency formatted ({currency_count}/49 cells) (0.10 pts)")
            total_score += 0.10
        elif currency_count >= 30:
            print(f"PARTIAL: Component 6 — D column partially formatted ({currency_count}/49 cells) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — Only {currency_count}/49 Balance cells have currency format")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Comment on G2 with 'Escalate to senior collector' (0.05 points)
    try:
        g2_comment = ws.cell(2, 7).comment
        if g2_comment is not None:
            comment_text = str(g2_comment.text).strip() if g2_comment.text else ''
            if 'escalate' in comment_text.lower() and 'senior' in comment_text.lower():
                print(f"PASS: Component 7 — G2 has comment: '{comment_text[:60]}' (0.05 pts)")
                total_score += 0.05
            else:
                print(f"FAIL: Component 7 — G2 comment exists but wrong text: '{comment_text[:60]}'")
        else:
            print(f"FAIL: Component 7 — No comment found on G2")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # Component 8: Summary row at row 52 with COUNTIF formulas (0.05 points)
    try:
        row52_a = ws.cell(52, 1).value
        row52_c = ws.cell(52, 3).value  # High COUNTIF
        row52_e = ws.cell(52, 5).value  # Medium COUNTIF
        row52_g = ws.cell(52, 7).value  # Low COUNTIF

        has_summary_label = row52_a is not None and 'priority' in str(row52_a).lower()
        has_high_countif = bool(row52_c and isinstance(row52_c, str) and
                               'COUNTIF' in row52_c.upper() and 'HIGH' in row52_c.upper())
        has_medium_countif = bool(row52_e and isinstance(row52_e, str) and
                                  'COUNTIF' in row52_e.upper() and 'MEDIUM' in row52_e.upper())
        has_low_countif = bool(row52_g and isinstance(row52_g, str) and
                               'COUNTIF' in row52_g.upper() and 'LOW' in row52_g.upper())

        countif_count = sum([has_high_countif, has_medium_countif, has_low_countif])
        if has_summary_label and countif_count >= 3:
            print(f"PASS: Component 8 — Summary row 52 with COUNTIF formulas (label='{row52_a}') (0.05 pts)")
            total_score += 0.05
        elif countif_count >= 2:
            print(f"PARTIAL: Component 8 — {countif_count}/3 COUNTIF formulas in row 52 (0.02 pts)")
            total_score += 0.02
        else:
            print(f"FAIL: Component 8 — Summary row 52 missing: label={has_summary_label}, COUNTIFs={countif_count}/3")
            print(f"  A52={repr(row52_a)}, C52={repr(row52_c)}, E52={repr(row52_e)}, G52={repr(row52_g)}")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
