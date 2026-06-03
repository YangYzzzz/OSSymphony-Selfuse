"""
Initial Setup: AR Collections prioritized call list
Task ID: calc_fin_ar_collections_064
Domain: libreoffice_calc

Creates an Accounts Receivable Collections worksheet with:
- Sheet 'ARCollections'
- Headers: Customer, Contact, Phone, Balance, Last Payment, Days Since Payment
- 49 rows of realistic customer data (rows 2-50)
- Column G intentionally empty (no Priority column yet)
- No sorting, no freeze panes, no conditional formatting, no comments
"""

import openpyxl
from openpyxl.styles import Font

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_ar_collections_064'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ARCollections'

    # --- Headers (Row 1) ---
    headers = ['Customer', 'Contact', 'Phone', 'Balance', 'Last Payment', 'Days Since Payment']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    # G1 intentionally left empty (no Priority header)

    # --- Realistic customer data (rows 2-50) ---
    # Columns: Customer, Contact, Phone, Balance, Last Payment, Days Since Payment
    customers = [
        ('Apex Manufacturing Inc', 'Robert Hines', '(312) 555-0142', 15420.75, '2025-10-05', 81),
        ('BlueStar Logistics LLC', 'Jennifer Walsh', '(415) 555-0289', 3280.50, '2025-11-18', 37),
        ('Cornerstone Retail Group', 'Michael Torres', '(713) 555-0374', 8940.00, '2025-10-22', 64),
        ('Delta Systems Corp', 'Amanda Pierce', '(617) 555-0461', 22100.30, '2025-09-15', 101),
        ('Echo Valley Supplies', 'Thomas Grant', '(206) 555-0558', 1850.20, '2025-11-28', 27),
        ('Frontier Tech Solutions', 'Sandra Nguyen', '(512) 555-0645', 6750.00, '2025-11-02', 53),
        ('Global Distribution Partners', 'Kevin Shaw', '(305) 555-0732', 12300.85, '2025-10-10', 76),
        ('Harbor Industrial Co', 'Linda Ortega', '(503) 555-0829', 4410.60, '2025-11-14', 41),
        ('Innovative Products Ltd', 'James Cooper', '(214) 555-0916', 19875.40, '2025-09-28', 88),
        ('Jefferson Wholesale Inc', 'Patricia Bell', '(404) 555-1003', 2640.15, '2025-11-22', 33),
        ('Keystone Materials LLC', 'Daniel Morgan', '(602) 555-1098', 9120.00, '2025-10-30', 56),
        ('Liberty Trading Co', 'Karen Rivera', '(816) 555-1185', 31500.00, '2025-09-08', 108),
        ('Midwest Auto Parts', 'Steven Hughes', '(314) 555-1272', 4875.30, '2025-11-10', 45),
        ('Nexus Healthcare Supplies', 'Nancy Brooks', '(216) 555-1369', 7340.50, '2025-10-25', 61),
        ('Olympia Foodservice Group', 'Charles Barnes', '(253) 555-1456', 1120.80, '2025-12-01', 24),
        ('Pacific Coast Distributors', 'Donna Ramirez', '(858) 555-1543', 18450.00, '2025-09-20', 96),
        ('Quality Control Systems', 'Matthew Cox', '(651) 555-1630', 5500.25, '2025-11-05', 50),
        ('Riverside Construction LLC', 'Helen Ward', '(916) 555-1727', 28900.00, '2025-09-02', 114),
        ('Summit Energy Services', 'Donald Peterson', '(720) 555-1814', 3960.70, '2025-11-20', 35),
        ('Titan Packaging Corp', 'Jessica Powell', '(901) 555-1901', 11200.00, '2025-10-15', 71),
        ('United Freight Services', 'Gary Stewart', '(502) 555-1998', 6230.45, '2025-11-08', 47),
        ('Vanguard Electronics', 'Ruth Hamilton', '(702) 555-2085', 24750.60, '2025-09-12', 104),
        ('Western Building Materials', 'Raymond Foster', '(503) 555-2172', 4100.00, '2025-11-16', 39),
        ('Xtreme Fitness Equipment', 'Christine Ross', '(480) 555-2269', 8600.80, '2025-10-28', 58),
        ('Yellowstone Resources Inc', 'Harold Sanchez', '(406) 555-2356', 2200.35, '2025-11-25', 30),
        ('Zenith Office Supplies', 'Deborah Morris', '(303) 555-2443', 13700.00, '2025-10-08', 78),
        ('Allstar Beverage Dist', 'Eugene Rogers', '(615) 555-2530', 5100.90, '2025-11-03', 52),
        ('Brightfield Farms LLC', 'Christine Reed', '(785) 555-2627', 1680.25, '2025-11-30', 25),
        ('Coastal Marine Supplies', 'Lawrence Cook', '(843) 555-2714', 9870.00, '2025-10-20', 66),
        ('Dynamo Electric Co', 'Judith Bailey', '(407) 555-2801', 16340.50, '2025-10-03', 83),
        ('Eastside Printing Group', 'Lawrence Bell', '(773) 555-2898', 3780.00, '2025-11-17', 38),
        ('Falcon Aerospace Parts', 'Carolyn Flores', '(321) 555-2985', 45600.00, '2025-08-25', 121),
        ('Green Valley Organics', 'Arthur Kelly', '(802) 555-3072', 2960.15, '2025-11-24', 31),
        ('Heritage Pharmacy Group', 'Shirley Sanders', '(918) 555-3169', 7890.30, '2025-10-26', 60),
        ('Interstate Trucking LLC', 'Jerry Price', '(931) 555-3256', 5750.00, '2025-11-06', 49),
        ('Jupiter Industrial Tools', 'Mildred Barnes', '(734) 555-3343', 10450.00, '2025-10-12', 74),
        ('Kessler Food Brokers', 'Ernest Henderson', '(972) 555-3430', 3330.75, '2025-11-19', 36),
        ('Landmark Office Products', 'Evelyn Coleman', '(314) 555-3527', 18900.00, '2025-09-18', 98),
        ('Metro Security Systems', 'Frederick Simmons', '(410) 555-3614', 6410.50, '2025-11-09', 46),
        ('Natural Stone Imports', 'Irene Washington', '(702) 555-3701', 4680.00, '2025-11-13', 42),
        ('Oakwood Furniture Dist', 'Carl Martin', '(601) 555-3798', 2100.60, '2025-11-27', 28),
        ('Premier Auto Glass', 'Wanda Thompson', '(520) 555-3885', 9450.00, '2025-10-23', 63),
        ('Quantum Data Services', 'Frank Garcia', '(512) 555-3972', 33800.00, '2025-08-30', 116),
        ('Raider Sports Equipment', 'Dorothy Martinez', '(702) 555-4059', 5200.40, '2025-11-04', 51),
        ('Suncoast Hospitality LLC', 'Ralph Anderson', '(813) 555-4146', 8120.75, '2025-10-27', 59),
        ('Triton Marine Exports', 'Phyllis Taylor', '(206) 555-4233', 14200.00, '2025-10-07', 79),
        ('Union Pacific Suppliers', 'Earl Thomas', '(402) 555-4320', 7030.20, '2025-10-29', 57),
        ('Vista Chemical Corp', 'Gladys Jackson', '(713) 555-4417', 25600.00, '2025-09-10', 106),
        ('Woodland Paper Products', 'Billy White', '(919) 555-4504', 4350.85, '2025-11-15', 40),
    ]

    for r, row_data in enumerate(customers, 2):
        ws.cell(row=r, column=1, value=row_data[0])   # Customer
        ws.cell(row=r, column=2, value=row_data[1])   # Contact
        ws.cell(row=r, column=3, value=row_data[2])   # Phone
        ws.cell(row=r, column=4, value=row_data[3])   # Balance
        ws.cell(row=r, column=5, value=row_data[4])   # Last Payment
        ws.cell(row=r, column=6, value=row_data[5])   # Days Since Payment
        # Column G intentionally left empty

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
