"""
Initial Setup: Create timesheet data for pivot table task
Task ID: calc_pivot_038
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_038'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

random.seed(42)

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Timesheet ---
    ws = wb.active
    ws.title = 'Timesheet'

    # 8 employees
    employees = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
        'Aisha Mohammed', 'Carlos Rivera', 'Emma Lindgren', 'Wei Zhang'
    ]

    # Projects and task types
    projects = [
        'Project Alpha', 'Project Beta', 'Project Gamma',
        'Project Delta', 'Infrastructure', 'Client Support'
    ]
    task_types = [
        'Development', 'Testing', 'Design', 'Documentation',
        'Meeting', 'Code Review', 'Planning', 'Deployment'
    ]

    # Generate dates: 01/01/2024 to 03/31/2024 = 91 days
    start_date = date(2024, 1, 1)
    end_date = date(2024, 3, 31)
    all_dates = []
    d = start_date
    while d <= end_date:
        all_dates.append(d)
        d += timedelta(days=1)

    # Headers
    headers = ['EntryID', 'Employee', 'Date', 'Hours', 'Project', 'TaskType']
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Generate 500 rows of timesheet data
    # Distribute entries so each employee gets roughly equal share across dates
    # Target: ~35-45 hours/week per employee, grand total ~4000
    # 13 weeks * 8 employees * 40 avg hours = 4160 total hours approximately
    # 500 entries, so average ~8 hours per entry

    entries = []
    entry_id = 1

    # Each employee works most weekdays
    for emp in employees:
        # Pick ~62-63 working dates for each employee (500/8 ≈ 62.5)
        # Filter to mostly weekdays
        weekdays = [d for d in all_dates if d.weekday() < 5]
        # Each employee works on about 62 of the ~65 weekdays
        work_dates = random.sample(weekdays, min(62, len(weekdays)))
        work_dates.sort()

        for wd in work_dates:
            hours = random.choice([6.0, 6.5, 7.0, 7.5, 8.0, 8.0, 8.0, 8.5, 9.0, 9.5, 10.0])
            project = random.choice(projects)
            task = random.choice(task_types)
            entries.append((entry_id, emp, wd, hours, project, task))
            entry_id += 1

    # Trim or pad to exactly 500
    if len(entries) > 500:
        entries = entries[:500]
    elif len(entries) < 500:
        while len(entries) < 500:
            emp = random.choice(employees)
            wd = random.choice(all_dates)
            hours = random.choice([6.0, 7.0, 7.5, 8.0, 8.5, 9.0])
            project = random.choice(projects)
            task = random.choice(task_types)
            entries.append((entry_id, emp, wd, hours, project, task))
            entry_id += 1

    # Sort by entry_id to keep sequential order and reassign IDs
    entries.sort(key=lambda x: (x[2], x[1]))  # sort by date, then employee
    for i in range(len(entries)):
        entries[i] = (i + 1,) + entries[i][1:]

    # Write data
    for r, (eid, emp, dt, hours, proj, task) in enumerate(entries, 2):
        ws.cell(row=r, column=1, value=eid)
        ws.cell(row=r, column=2, value=emp)
        ws.cell(row=r, column=3, value=dt)
        ws.cell(row=r, column=3).number_format = 'MM/DD/YYYY'
        ws.cell(row=r, column=4, value=hours)
        ws.cell(row=r, column=4).number_format = '0.0'
        ws.cell(row=r, column=5, value=proj)
        ws.cell(row=r, column=6, value=task)

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 16

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Count total hours for verification
    total_hours = sum(e[3] for e in entries)
    print(f'Total entries: {len(entries)}, Total hours: {total_hours}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
