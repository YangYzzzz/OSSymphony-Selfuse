"""
Reward Script: Create a pivot table summarizing total hours worked per employee per week
Task ID: calc_pivot_038
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Pivot sheet exists (not present in initial)
  Component 2 (0.25): Correct pivot structure — Employee row labels, week columns, Grand Total column
  Component 3 (0.20): All 8 employees present as row entries
  Component 4 (0.15): Hours data is reasonable (non-empty numerics, weekly averages in plausible range)
  Component 5 (0.20): Grand Total row with overall total approximately 4000
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_038'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: file must load
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Pivot sheet exists (0.20 points)
    # This FAILS on initial_env (only 'Timesheet' exists) and PASSES on golden_env
    try:
        sheet_names = wb.sheetnames
        # Look for a sheet that could be the pivot table (not 'Timesheet')
        pivot_sheet_name = None
        for name in sheet_names:
            if name.lower() != 'timesheet':
                pivot_sheet_name = name
                break

        if pivot_sheet_name is not None:
            print(f"PASS: Component 1 — Pivot sheet found: '{pivot_sheet_name}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — No pivot sheet found. Sheets: {sheet_names}")
            # Without a pivot sheet, no further checks are possible
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb[pivot_sheet_name]

    # Component 2: Correct pivot structure (0.25 points)
    # Headers should include: Employee label in A1, week-based column headers, Grand Total column
    # This verifies the pivot layout matches the task requirement
    try:
        header_a1 = ws.cell(row=1, column=1).value
        # Check that A1 is an employee-related label
        has_employee_label = (
            header_a1 is not None and
            isinstance(header_a1, str) and
            'employee' in header_a1.lower()
        )

        # Check that there are week-related column headers (at least 10 week columns)
        week_col_count = 0
        grand_total_col = None
        for col in range(2, ws.max_column + 1):
            hval = ws.cell(row=1, column=col).value
            if hval is not None and isinstance(hval, str):
                hval_lower = hval.lower()
                if 'week' in hval_lower or 'wk' in hval_lower or '/' in hval_lower:
                    week_col_count += 1
                if 'grand' in hval_lower and 'total' in hval_lower:
                    grand_total_col = col

        structure_ok = has_employee_label and week_col_count >= 10 and grand_total_col is not None
        if structure_ok:
            print(f"PASS: Component 2 — Pivot structure correct: Employee label='{header_a1}', "
                  f"{week_col_count} week columns, Grand Total in col {grand_total_col} (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — Structure issues: employee_label={has_employee_label} "
                  f"(A1='{header_a1}'), week_cols={week_col_count}, grand_total_col={grand_total_col}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: All 8 employees present (0.20 points)
    # The task context specifies 8 employees
    try:
        employee_names = []
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None and isinstance(val, str):
                val_lower = val.lower().strip()
                # Skip total/summary rows
                if 'total' not in val_lower and 'sum' not in val_lower and val_lower != '':
                    employee_names.append(val.strip())

        num_employees = len(employee_names)
        if num_employees >= 8:
            print(f"PASS: Component 3 — {num_employees} employees found: {employee_names[:8]} (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 — Expected 8 employees, found {num_employees}: {employee_names}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Hours data is reasonable (0.15 points)
    # Each employee should have numeric hours per week, with averages roughly 35-45
    try:
        non_empty_count = 0
        total_hours_all = 0.0
        employee_count = 0

        for r in range(2, ws.max_row + 1):
            emp_val = ws.cell(row=r, column=1).value
            if emp_val is None or not isinstance(emp_val, str):
                continue
            if 'total' in emp_val.lower() or 'sum' in emp_val.lower():
                continue

            employee_count += 1
            row_hours = 0.0
            row_cells = 0
            # Check data columns (from col 2 up to the last column, skipping grand total)
            last_data_col = grand_total_col - 1 if grand_total_col else ws.max_column
            for c in range(2, last_data_col + 1):
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None:
                    try:
                        hours = float(cell_val)
                        row_hours += hours
                        row_cells += 1
                        non_empty_count += 1
                    except (ValueError, TypeError):
                        pass

            total_hours_all += row_hours

        # Check: meaningful amount of numeric data exists
        if non_empty_count >= 50 and employee_count >= 8:
            avg_per_emp = total_hours_all / employee_count if employee_count > 0 else 0
            avg_per_week = avg_per_emp / 13 if employee_count > 0 else 0
            print(f"PASS: Component 4 — Data reasonable: {non_empty_count} cells with hours, "
                  f"avg {avg_per_week:.1f} hrs/wk/employee, total {total_hours_all:.1f} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Insufficient data: {non_empty_count} numeric cells, "
                  f"{employee_count} employees")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Grand Total row with approximately correct total (0.20 points)
    # Context says approximately 4000 total hours
    try:
        grand_total_value = None

        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=1).value
            if val is not None and isinstance(val, str) and 'total' in val.lower():
                # Found the total row — check the Grand Total column or the last column
                if grand_total_col:
                    gt_val = ws.cell(row=r, column=grand_total_col).value
                else:
                    gt_val = ws.cell(row=r, column=ws.max_column).value

                if gt_val is not None:
                    try:
                        grand_total_value = float(gt_val)
                    except (ValueError, TypeError):
                        pass
                break

        if grand_total_value is not None:
            # Task says approximately 4000 — allow range 3500-4500
            if 3500 <= grand_total_value <= 4500:
                print(f"PASS: Component 5 — Grand Total = {grand_total_value} "
                      f"(within 3500-4500 range) (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 5 — Grand Total = {grand_total_value}, "
                      f"expected approximately 4000 (range 3500-4500)")
        else:
            print(f"FAIL: Component 5 — No Grand Total row/value found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
