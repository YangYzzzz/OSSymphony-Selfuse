"""
Reward Script: Apply whole number validation to cell C2 (1-100 inclusive)
Task ID: calc_nrv_047
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35) - Data validation exists on C2 with type=whole
  Component 2 (0.35) - Validation uses between operator with min=1, max=100
  Component 3 (0.30) - Error message is configured (showErrorMessage=True)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_047'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Precondition: headers should still be intact
    if ws['A1'].value != 'Student' or ws['B1'].value != 'Subject' or ws['C1'].value != 'Score':
        print("WARN: Header row changed, but continuing with validation checks")

    # Find data validation targeting C2
    dv_on_c2 = None
    for dv in ws.data_validations.dataValidation:
        sqref_str = str(dv.sqref)
        # Check if C2 is covered by this validation's range
        if 'C2' in sqref_str:
            dv_on_c2 = dv
            break

    if dv_on_c2 is None:
        print("FAIL: No data validation found on cell C2")
        print(f"  Total validations in sheet: {len(ws.data_validations.dataValidation)}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Data validation exists on C2 with type=whole (0.35 points)
    try:
        if dv_on_c2.type == 'whole':
            print(f"PASS: Component 1 - Data validation type is 'whole' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 - Expected type='whole', found type='{dv_on_c2.type}'")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Validation uses 'between' operator with min=1, max=100 (0.35 points)
    try:
        operator_ok = (dv_on_c2.operator == 'between')
        min_ok = (str(dv_on_c2.formula1).strip() == '1')
        max_ok = (str(dv_on_c2.formula2).strip() == '100')

        if operator_ok and min_ok and max_ok:
            print(f"PASS: Component 2 - Operator='between', min=1, max=100 (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 - operator={dv_on_c2.operator}, formula1={dv_on_c2.formula1}, formula2={dv_on_c2.formula2}")
            # Partial credit: correct operator but wrong bounds
            if operator_ok:
                print(f"  Partial: operator is correct but bounds are wrong")
                total_score += 0.1
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: Error message is enabled (showErrorMessage=True) (0.30 points)
    try:
        if dv_on_c2.showErrorMessage:
            print(f"PASS: Component 3 - showErrorMessage is True (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 - showErrorMessage is {dv_on_c2.showErrorMessage}")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
