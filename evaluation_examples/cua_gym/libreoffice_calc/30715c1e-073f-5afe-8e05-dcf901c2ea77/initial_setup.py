"""
Initial Setup: Apply whole number validation to cell C2 (1-100)
Task ID: calc_nrv_047
Domain: libreoffice_calc

Creates a spreadsheet with Student/Subject/Score headers and realistic
student data. C2 is empty with NO validation — the agent must add it.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_047'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Headers
    ws["A1"] = "Student"
    ws["B1"] = "Subject"
    ws["C1"] = "Score"

    # Realistic student data (C2 intentionally left empty — agent task target)
    data = [
        ["Emma Rodriguez", "Mathematics", None],       # Row 2: C2 empty (task target)
        ["Liam Nakamura", "Physics", 88],               # Row 3
        ["Sophia Patel", "Chemistry", 76],              # Row 4
        ["Noah Kim", "Biology", 92],                    # Row 5
        ["Olivia Johansson", "Mathematics", 67],        # Row 6
        ["Ethan Okafor", "English Literature", 81],     # Row 7
        ["Ava Petrov", "History", 95],                  # Row 8
        ["Mason Al-Farsi", "Computer Science", 73],     # Row 9
        ["Isabella Chen", "Physics", 84],               # Row 10
        ["James Morales", "Chemistry", 59],             # Row 11
        ["Charlotte Dubois", "Biology", 90],            # Row 12
        ["Benjamin Tanaka", "English Literature", 77],  # Row 13
    ]

    for r, row_data in enumerate(data, 2):
        ws.cell(row=r, column=1, value=row_data[0])
        ws.cell(row=r, column=2, value=row_data[1])
        if row_data[2] is not None:
            ws.cell(row=r, column=3, value=row_data[2])

    # Adjust column widths for readability
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
