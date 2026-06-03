"""
Initial Setup: Workbook with 4 sheets for team review preparation
Task ID: calc_sht_multiop_002
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_multiop_002'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Scratch (temporary calculations, empty) ---
    ws_scratch = wb.active
    ws_scratch.title = 'Scratch'
    # Leave scratch empty as per task context

    # --- Sheet 2: Sales Report (main report, 300 rows, row 1 headers) ---
    ws_sales = wb.create_sheet('Sales Report')

    # Headers in row 1
    headers = ['Date', 'Region', 'Salesperson', 'Product', 'Units Sold', 'Unit Price', 'Revenue', 'Discount', 'Net Revenue', 'Status']
    for col, h in enumerate(headers, 1):
        ws_sales.cell(row=1, column=col, value=h)

    # Realistic data — 300 rows
    regions = ['North', 'South', 'East', 'West', 'Central']
    salespeople = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
        'Laura Patel', 'James Okafor', 'Sophie Williams', 'Ryan Nguyen',
        'Aisha Thompson', 'Carlos Martinez', 'Nina Petrov', 'Derek Brown',
        'Isabel Santos', 'Kevin O\'Brien', 'Mei Liu', 'Thomas Anderson'
    ]
    products = [
        'Enterprise Suite Pro', 'Analytics Dashboard', 'Cloud Storage 1TB',
        'Security Firewall', 'Data Pipeline Tool', 'Collaboration Hub',
        'Mobile SDK License', 'API Gateway Access', 'ML Model Training',
        'Support Premium Plan'
    ]
    statuses = ['Closed', 'Pending', 'In Progress', 'Cancelled', 'Closed']

    import datetime
    base_date = datetime.date(2024, 1, 2)

    for row in range(2, 302):
        i = row - 2
        date = base_date + datetime.timedelta(days=i)
        region = regions[i % len(regions)]
        salesperson = salespeople[i % len(salespeople)]
        product = products[i % len(products)]
        units = 5 + (i * 7 % 46)
        unit_price = round(150.00 + (i * 13.7 % 850), 2)
        revenue = round(units * unit_price, 2)
        discount = round(0.05 + (i % 10) * 0.01, 2)
        net_revenue = round(revenue * (1 - discount), 2)
        status = statuses[i % len(statuses)]

        ws_sales.cell(row=row, column=1, value=date.strftime('%Y-%m-%d'))
        ws_sales.cell(row=row, column=2, value=region)
        ws_sales.cell(row=row, column=3, value=salesperson)
        ws_sales.cell(row=row, column=4, value=product)
        ws_sales.cell(row=row, column=5, value=units)
        ws_sales.cell(row=row, column=6, value=unit_price)
        ws_sales.cell(row=row, column=7, value=revenue)
        ws_sales.cell(row=row, column=8, value=discount)
        ws_sales.cell(row=row, column=9, value=net_revenue)
        ws_sales.cell(row=row, column=10, value=status)

    # NO freeze panes — initial state (task requires adding freeze)

    # --- Sheet 3: Cost Analysis ---
    ws_cost = wb.create_sheet('Cost Analysis')

    cost_headers = ['Product', 'Category', 'Unit Cost', 'Volume', 'Total Cost', 'Margin %', 'Quarter']
    for col, h in enumerate(cost_headers, 1):
        ws_cost.cell(row=1, column=col, value=h)

    cost_categories = ['Software', 'Hardware', 'Services', 'Licensing']
    quarters = ['Q1 2024', 'Q2 2024', 'Q3 2024', 'Q4 2024']
    for row in range(2, 52):
        i = row - 2
        product = products[i % len(products)]
        category = cost_categories[i % len(cost_categories)]
        unit_cost = round(45.00 + (i * 8.3 % 320), 2)
        volume = 10 + (i * 5 % 90)
        total_cost = round(unit_cost * volume, 2)
        margin = round(0.20 + (i % 7) * 0.05, 2)
        quarter = quarters[i % len(quarters)]

        ws_cost.cell(row=row, column=1, value=product)
        ws_cost.cell(row=row, column=2, value=category)
        ws_cost.cell(row=row, column=3, value=unit_cost)
        ws_cost.cell(row=row, column=4, value=volume)
        ws_cost.cell(row=row, column=5, value=total_cost)
        ws_cost.cell(row=row, column=6, value=margin)
        ws_cost.cell(row=row, column=7, value=quarter)

    # --- Sheet 4: Summary ---
    ws_summary = wb.create_sheet('Summary')

    ws_summary.cell(row=1, column=1, value='Metric')
    ws_summary.cell(row=1, column=2, value='Q1 2024')
    ws_summary.cell(row=1, column=3, value='Q2 2024')
    ws_summary.cell(row=1, column=4, value='Q3 2024')
    ws_summary.cell(row=1, column=5, value='Q4 2024')
    ws_summary.cell(row=1, column=6, value='Annual Total')

    summary_data = [
        ['Total Revenue', 245320.50, 278940.75, 301560.20, 324890.30, 1150711.75],
        ['Total Units Sold', 1240, 1385, 1502, 1628, 5755],
        ['Avg Deal Size', 197.84, 201.40, 200.77, 199.56, 199.95],
        ['New Customers', 48, 55, 61, 67, 231],
        ['Renewal Rate %', 0.87, 0.89, 0.91, 0.92, 0.90],
        ['Net Revenue', 213578.84, 242878.45, 262457.37, 282654.56, 1001569.22],
        ['Cost of Sales', 98128.20, 111576.30, 120624.08, 129956.12, 460284.70],
        ['Gross Profit', 115450.64, 131302.15, 141833.29, 152698.44, 541284.52],
        ['Gross Margin %', 0.54, 0.54, 0.54, 0.54, 0.54],
        ['Top Region', 'West', 'North', 'West', 'East', 'West'],
    ]

    for r, row_data in enumerate(summary_data, 2):
        for c, val in enumerate(row_data, 1):
            ws_summary.cell(row=r, column=c, value=val)

    # Verify sheet order: Scratch, Sales Report, Cost Analysis, Summary
    assert wb.sheetnames == ['Scratch', 'Sales Report', 'Cost Analysis', 'Summary'], \
        f"Sheet order mismatch: {wb.sheetnames}"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheets: {wb.sheetnames}')
    print(f'Sales Report rows: {ws_sales.max_row}')
    print(f'No freeze panes on Sales Report: {ws_sales.freeze_panes}')
    print(f'Scratch visible: {ws_scratch.sheet_state}')


create_initial()
