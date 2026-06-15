"""
Reward Script: Set up workbook for team review
Task ID: calc_sht_multiop_002
Domain: libreoffice_calc

Scoring Rubric:
  Component 1: 'Sales Report' is the first sheet in the workbook (0.35 pts)
  Component 2: 'Sales Report' has freeze panes at A2 (row 1 frozen)     (0.35 pts)
  Component 3: 'Scratch' sheet is hidden                                 (0.30 pts)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sht_multiop_002'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Set up this workbook for a team review:
      - Freeze row 1 in the 'Sales Report' sheet
      - Hide the 'Scratch' sheet
      - Move 'Sales Report' to be the first sheet
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist as a precondition gate
    required_sheets = {'Sales Report', 'Scratch'}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        print(f"CRITICAL: Missing required sheets: {missing}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: 'Sales Report' is the first sheet in the workbook (0.35 points)
    # In the initial file, 'Scratch' is first and 'Sales Report' is second.
    # This component FAILS on initial -> PASSES on golden.
    try:
        first_sheet_name = wb.sheetnames[0]
        if first_sheet_name == 'Sales Report':
            print(f"PASS: Component 1 — 'Sales Report' is the first sheet (index 0) (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Expected 'Sales Report' as first sheet, found: '{first_sheet_name}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: 'Sales Report' has freeze panes at A2 (row 1 frozen) (0.35 points)
    # In the initial file, freeze_panes is None (no freeze applied).
    # This component FAILS on initial -> PASSES on golden.
    try:
        ws_sales = wb['Sales Report']
        freeze = ws_sales.freeze_panes
        if freeze == 'A2':
            print(f"PASS: Component 2 — 'Sales Report' freeze_panes == 'A2' (row 1 frozen) (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 — Expected freeze_panes='A2', found: {repr(freeze)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: 'Scratch' sheet is hidden (0.30 points)
    # In the initial file, 'Scratch' is visible.
    # This component FAILS on initial -> PASSES on golden.
    try:
        ws_scratch = wb['Scratch']
        state = ws_scratch.sheet_state
        if state == 'hidden':
            print(f"PASS: Component 3 — 'Scratch' sheet is hidden (sheet_state='hidden') (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — Expected 'Scratch' sheet_state='hidden', found: '{state}'")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
