"""
FINAL REWARD SCRIPT - SUCCESS
Task: I want to copy 'Badge ID' to 'Security Badge' and pad with zeros on the left to reach 5 digits for each entry. Complete this and leave irrelevant regions unchanged.
Generated: 2025-11-24 07:25:45
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl
import os
import pathlib


def verify_security_badges(file_path: str) -> float:
    """Verify that the workbook satisfies the following:
    1. Column 'Security Badge' contains the value of 'Badge ID' padded with
       leading zeros to 5 digits (e.g. 12 → 00012).
    2. All other columns/values remain unchanged from the expected initial
       state (Employee Name, Department, Badge ID).

    A progressive score is returned:
      • 0.7 weight for correct Security Badge values.
      • 0.3 weight for keeping other data unchanged.
    The final score is capped at 1.0 and rounded to 4 decimals.
    """
    print(f"Verifying workbook at: {file_path}")

    # Weighting configuration
    security_weight = 0.7   # 70 % of score for correct Security Badge column
    unchanged_weight = 0.3  # 30 % of score for leaving other data untouched

    # Expected reference data (initial state):
    expected_rows = [
        ("Alice Smith",   "IT",      "12"  ),
        ("Bob Johnson",   "HR",      "345" ),
        ("Charlie Davis", "Finance", "7890"),
        ("Dana Lee",      "IT",      "56"  ),
        ("Evan Grant",    "Ops",     "4"   ),
    ]

    # Helper: normalise any cell value to string for comparison
    def norm(value):
        if value is None:
            return ""
        # Convert numerics to int if whole-number, then to str
        if isinstance(value, (int, float)):
            if abs(value - int(value)) < 1e-9:
                value = int(value)
            return str(value)
        return str(value).strip()

    # ------------------------------------------------------------
    # Load workbook & locate sheet
    # ------------------------------------------------------------
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        return 0.0

    if "EmployeeData" not in wb.sheetnames:
        print("✗ Sheet 'EmployeeData' not found")
        return 0.0
    sheet = wb["EmployeeData"]

    # ------------------------------------------------------------
    # Identify header indices
    # ------------------------------------------------------------
    header = [norm(c.value) for c in sheet[1]]
    try:
        name_idx = header.index("Employee Name")
        dept_idx = header.index("Department")
        badge_idx = header.index("Badge ID")
        sec_idx = header.index("Security Badge")
    except ValueError as e:
        print(f"✗ Required header missing: {e}")
        return 0.0

    # ------------------------------------------------------------
    # Gather data rows (starting row 2)
    # ------------------------------------------------------------
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    if not data_rows:
        print("✗ No data rows found")
        return 0.0

    # ------------------------------------------------------------
    # 1) Verify Security Badge column
    # ------------------------------------------------------------
    correct_sec = 0
    total_sec  = len(expected_rows)  # expected meaningful rows

    for r_idx, row in enumerate(data_rows, start=2):
        if r_idx - 2 >= total_sec:
            break  # Ignore potential extra rows

        badge_raw = norm(row[badge_idx])
        sec_raw   = norm(row[sec_idx])

        # Skip fully blank rows
        if badge_raw == "" and sec_raw == "":
            continue

        # Determine expected padded value
        try:
            expected_sec = f"{int(float(badge_raw)):05d}"
        except ValueError:
            expected_sec = badge_raw.zfill(5)[:5]

        if sec_raw == expected_sec:
            correct_sec += 1
        else:
            print(f"Row {r_idx}: Expected Security Badge '{expected_sec}' but found '{sec_raw}'")

    security_score = security_weight * (correct_sec / total_sec)
    print(f"Security badge correctness: {correct_sec}/{total_sec} -> {security_score:.3f}")

    # ------------------------------------------------------------
    # 2) Verify unchanged data regions (Employee Name, Department, Badge ID)
    # ------------------------------------------------------------
    unchanged_total   = 0
    unchanged_correct = 0

    for idx, expected in enumerate(expected_rows):
        if idx >= len(data_rows):
            print(f"Row {idx+2}: Missing expected row data")
            continue
        row = data_rows[idx]
        actual_values = (
            norm(row[name_idx]),
            norm(row[dept_idx]),
            norm(row[badge_idx]),
        )
        for field_name, act, exp in zip(["Name", "Dept", "Badge"], actual_values, expected):
            unchanged_total += 1
            if act == exp:
                unchanged_correct += 1
            else:
                print(f"Row {idx+2}: {field_name} changed. Expected '{exp}' but found '{act}'")

    unchanged_score = (
        unchanged_weight * (unchanged_correct / unchanged_total)
        if unchanged_total else 0.0
    )
    print(f"Unchanged data correctness: {unchanged_correct}/{unchanged_total} -> {unchanged_score:.3f}")

    # ------------------------------------------------------------
    # Combine scores, cap at 1.0, round, and return
    # ------------------------------------------------------------
    final_score = round(min(security_score + unchanged_score, 1.0), 4)
    print(f"Final computed score: {final_score:.3f}")
    return final_score


# ------------------------------------------------------------------
# Convenience helper to locate likely workbook in /home/user
# ------------------------------------------------------------------

def find_target_file() -> str | None:
    search_root = pathlib.Path("/home/user")
    patterns = ("*.xlsx", "*.xlsm")
    candidates = [p for pattern in patterns for p in search_root.glob(pattern)]
    if not candidates:
        return None
    # Heuristic: filenames containing these keywords get higher priority
    keywords = ["badge", "security", "zeros", "pad"]
    candidates.sort(key=lambda p: sum(k in p.name.lower() for k in keywords), reverse=True)
    return str(candidates[0])


# ------------------------------------------------------------------
# Main entry point – locate file, run verification, print REWARD
# ------------------------------------------------------------------

def main():
    default_path = "/home/user/i_want_to_copy_badge_id_to_security_badge_and_pad_with_zeros_on_the_left_to_reach_5_digits_for_each_.xlsx"
    target = default_path if os.path.exists(default_path) else find_target_file()

    if not target or not os.path.exists(target):
        print("✗ Target Excel file not found for verification.")
        print("REWARD: 0.0")
        return

    reward = verify_security_badges(target)
    print(f"REWARD: {reward}")


if __name__ == "__main__":
    main()
