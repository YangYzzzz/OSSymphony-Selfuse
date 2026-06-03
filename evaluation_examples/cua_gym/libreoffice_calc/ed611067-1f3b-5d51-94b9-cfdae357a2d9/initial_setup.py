"""
Initial Setup: Sales Commission Calculator (Tiered)
Task ID: calc_sales_commission_tiered_005
Domain: libreoffice_calc

Creates a Commissions sheet with 15 sales reps.
Columns D (Attainment %), E (Commission Rate), F (Commission Earned) are EMPTY
— the agent must fill them in with formulas.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_commission_tiered_005'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Commissions'

    # --- Headers ---
    headers = ['Rep Name', 'Quota', 'Actual Sales', 'Attainment %', 'Commission Rate', 'Commission Earned']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, name='Calibri', size=11, color='FFFFFF')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # --- Column widths ---
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

    # --- Data: 15 sales reps ---
    # Quota range: $600,000 - $1,500,000
    # Actual Sales range: $280,000 - $1,800,000
    # NOTE: Columns D, E, F intentionally left EMPTY
    data = [
        # Rep Name,          Quota,     Actual Sales
        ('Diana Morales',    900000,    450000),    # ~50% attainment -> 5% tier
        ('James Thornton',   750000,    625000),    # ~83% -> 10% tier
        ('Priya Kapoor',     1200000,   1320000),   # 110% -> 12% tier (exceeded)
        ('Marcus Webb',      800000,    396000),    # ~49.5% -> 5% tier
        ('Sofia Lindqvist',  1000000,   1050000),   # 105% -> 12% tier (exceeded)
        ('Andre Fontaine',   600000,    504000),    # 84% -> 10% tier
        ('Yuki Tanaka',      1100000,   858000),    # ~78% -> 8% tier
        ('Rachel Okonkwo',   950000,    1045000),   # 110% -> 12% tier (exceeded)
        ('Derek Simmons',    700000,    518000),    # ~74% -> 8% tier
        ('Mei-Ling Chen',    1500000,   1800000),   # 120% -> 12% tier (exceeded)
        ('Gabriel Torres',   850000,    680000),    # 80% -> 10% tier (boundary)
        ('Natasha Ivanova',  650000,    305000),    # ~47% -> 5% tier
        ('Omar Khalil',      1300000,   1391000),   # 107% -> 12% tier (exceeded)
        ('Bridget Sullivan', 1050000,   756000),    # ~72% -> 8% tier
        ('Ravi Patel',       1150000,   1196000),   # ~104% -> 12% tier (exceeded)
    ]

    for r, (name, quota, actual) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=name)
        quota_cell = ws.cell(row=r, column=2, value=quota)
        quota_cell.number_format = '$#,##0'
        actual_cell = ws.cell(row=r, column=3, value=actual)
        actual_cell.number_format = '$#,##0'
        # Columns D, E, F intentionally left EMPTY

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Commissions')
    print('Rows: 15 sales reps (rows 2-16)')
    print('Columns D, E, F are EMPTY — agent must add formulas')


create_initial()
