"""
Reward Script: Build a tiered commission calculator with IFS formula and gold fill for quota-exceeding reps.
Task ID: calc_sales_commission_tiered_005
Domain: libreoffice_calc
Scoring:
  Component 1: Attainment % formulas in D2:D16 (=Cx/Bx pattern)         — 0.25 pts
  Component 2: IFS commission rate formulas in E2:E16 (tiered 5/8/10/12%) — 0.35 pts
  Component 3: Commission earned formulas in F2:F16 (=Cx*Ex pattern)      — 0.20 pts
  Component 4: Conditional formatting gold/yellow fill for D >= 100%       — 0.20 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_commission_tiered_005'


def normalize_formula(f):
    """Normalize formula string: upper-case, remove spaces."""
    if not isinstance(f, str):
        return ''
    return f.upper().replace(' ', '')


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Commissions' sheet must exist
    if 'Commissions' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Commissions' not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Commissions']

    # -----------------------------------------------------------------------
    # Component 1: Attainment % formulas in D2:D16 (0.25 points)
    # Each cell Dx should contain a formula =Cx/Bx (attainment = actual/quota)
    # This FAILS on initial (all None) and PASSES on golden (all formulas).
    # -----------------------------------------------------------------------
    try:
        attainment_ok = 0
        attainment_total = 15  # rows 2..16
        for row in range(2, 17):
            val = ws.cell(row=row, column=4).value  # column D
            if val is None:
                continue
            norm = normalize_formula(str(val))
            # Accept =Cx/Bx pattern  e.g. =C2/B2
            expected = f'=C{row}/B{row}'
            if norm == normalize_formula(expected):
                attainment_ok += 1

        if attainment_ok == attainment_total:
            print(f"PASS: Component 1 — All {attainment_total} attainment formulas present in D2:D16 (0.25 pts)")
            total_score += 0.25
        elif attainment_ok >= attainment_total // 2:
            # Partial credit: at least half correct
            partial = round(0.25 * attainment_ok / attainment_total, 4)
            print(f"PARTIAL: Component 1 — {attainment_ok}/{attainment_total} attainment formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {attainment_ok}/{attainment_total} attainment formulas in D2:D16")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: IFS commission rate formulas in E2:E16 (0.35 points)
    # Each cell Ex should contain an IFS formula implementing the four tiers:
    #   <=50% -> 5%, <=80% -> 8%, <=100% -> 10%, >100% -> 12%
    # The formula must reference the corresponding Dx cell.
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        ifs_ok = 0
        ifs_total = 15
        for row in range(2, 17):
            val = ws.cell(row=row, column=5).value  # column E
            if val is None:
                continue
            norm = normalize_formula(str(val))
            # Must start with =IFS(
            if not norm.startswith('=IFS('):
                continue
            # Must contain the four rate values: 0.05, 0.08, 0.10/0.1, 0.12
            has_5pct = '0.05' in norm
            has_8pct = '0.08' in norm
            has_10pct = ('0.10' in norm or ',0.1,' in norm or norm.endswith(',0.1)'))
            has_12pct = '0.12' in norm
            # Must reference the correct Dx cell for this row
            d_ref = f'D{row}'
            has_d_ref = d_ref in norm.upper()
            if has_5pct and has_8pct and has_10pct and has_12pct and has_d_ref:
                ifs_ok += 1

        if ifs_ok == ifs_total:
            print(f"PASS: Component 2 — All {ifs_total} IFS commission rate formulas present in E2:E16 (0.35 pts)")
            total_score += 0.35
        elif ifs_ok >= ifs_total // 2:
            partial = round(0.35 * ifs_ok / ifs_total, 4)
            print(f"PARTIAL: Component 2 — {ifs_ok}/{ifs_total} IFS formulas with correct tiers ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Only {ifs_ok}/{ifs_total} valid IFS formulas in E2:E16")
            # Provide diagnostic for first non-None entry
            for row in range(2, 17):
                val = ws.cell(row=row, column=5).value
                if val is not None:
                    print(f"  Sample E{row}: {repr(val)}")
                    break
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -----------------------------------------------------------------------
    # Component 3: Commission earned formulas in F2:F16 (0.20 points)
    # Each cell Fx should contain a formula =Cx*Ex (actual_sales * rate)
    # This FAILS on initial (all None) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        earned_ok = 0
        earned_total = 15
        for row in range(2, 17):
            val = ws.cell(row=row, column=6).value  # column F
            if val is None:
                continue
            norm = normalize_formula(str(val))
            expected = f'=C{row}*E{row}'
            if norm == normalize_formula(expected):
                earned_ok += 1

        if earned_ok == earned_total:
            print(f"PASS: Component 3 — All {earned_total} commission earned formulas present in F2:F16 (0.20 pts)")
            total_score += 0.20
        elif earned_ok >= earned_total // 2:
            partial = round(0.20 * earned_ok / earned_total, 4)
            print(f"PARTIAL: Component 3 — {earned_ok}/{earned_total} commission earned formulas correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {earned_ok}/{earned_total} commission earned formulas in F2:F16")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -----------------------------------------------------------------------
    # Component 4: Conditional formatting with gold/yellow fill for D >= 100% (0.20 points)
    # A formula-based rule should apply a gold/yellow fill to rows where attainment >= 1
    # (i.e. $D2>=1 or similar). The fill color should be gold/yellow (ARGB approx FFFFD700).
    # This FAILS on initial (no CF rules) and PASSES on golden.
    # -----------------------------------------------------------------------
    try:
        cf_found = False
        cf_gold_fill = False

        # Gold/yellow ARGB variants to accept:
        GOLD_COLORS = {
            'FFFFD700',  # pure gold
            'FFFFFF00',  # pure yellow
            'FFFFCC00',  # amber
            'FFFFC000',  # dark yellow/gold (Office default)
            'FFFFE599',  # light gold
            'FFFFD966',  # Office gold
        }

        cf_rules = ws.conditional_formatting
        for cf_range in cf_rules:
            for rule in cf_rules[cf_range]:
                # Must be a formula-based rule or CellIs rule referencing column D
                rule_type = getattr(rule, 'type', '')
                formulas = getattr(rule, 'formula', []) or []

                is_d_gte_1 = False
                if rule_type == 'expression' and formulas:
                    formula_str = normalize_formula(str(formulas[0]))
                    # Accept patterns: $D2>=1, D2>=1, $D2>=1.0, etc.
                    if re.search(r'\$?D\d+>=1', formula_str):
                        is_d_gte_1 = True
                    # Also accept formulas checking D >= 100% expressed as decimal
                    elif re.search(r'\$?D\d+>=0\.?1{2}', formula_str):
                        is_d_gte_1 = True

                if is_d_gte_1:
                    cf_found = True
                    # Check for gold/yellow fill in the differential format
                    try:
                        dxf = rule.dxf
                        if dxf and dxf.fill:
                            fg_color = dxf.fill.fgColor.rgb
                            # Accept if color matches any gold/yellow or starts with FF and
                            # has high R, moderate G, low B (yellow-ish family)
                            if fg_color in GOLD_COLORS:
                                cf_gold_fill = True
                            else:
                                # Broader check: yellow-family (R=FF, G>=99, B<=66 in hex)
                                try:
                                    r_val = int(fg_color[2:4], 16)
                                    g_val = int(fg_color[4:6], 16)
                                    b_val = int(fg_color[6:8], 16)
                                    if r_val >= 200 and g_val >= 150 and b_val <= 100:
                                        cf_gold_fill = True
                                except Exception:
                                    pass
                            if not cf_gold_fill:
                                print(f"  CF rule found but fill color is {fg_color} (expected gold/yellow)")
                    except Exception as fill_e:
                        print(f"  CF fill check error: {fill_e}")

        if cf_found and cf_gold_fill:
            print("PASS: Component 4 — Conditional formatting with gold fill for quota-exceeding rows (0.20 pts)")
            total_score += 0.20
        elif cf_found:
            # Rule exists but fill color doesn't match — give partial credit
            print("PARTIAL: Component 4 — CF rule for D>=1 found but fill color is not gold/yellow (0.10 pts)")
            total_score += 0.10
        else:
            print("FAIL: Component 4 — No conditional formatting rule found for quota-exceeding rows (D>=100%)")
            # Diagnostic: show any CF rules present
            rule_count = sum(1 for _ in cf_rules)
            print(f"  Total CF ranges: {rule_count}")
            for cf_range in cf_rules:
                for rule in cf_rules[cf_range]:
                    fmls = getattr(rule, 'formula', [])
                    print(f"  CF: range={cf_range}, type={getattr(rule,'type','?')}, formula={fmls}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
