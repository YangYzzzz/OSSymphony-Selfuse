"""
Initial Setup: Patient intake Excel spreadsheet and PDF template
Task ID: osworld_multi_apps_excel_pdf_form_003
Domain: libreoffice_calc (multi-app: Calc + PDF template)

Creates:
  - /home/user/patient_intake.xlsx  (patient data spreadsheet)
  - /home/user/Desktop/intake_form_template.pdf  (blank intake form template)

Opens:
  - LibreOffice Calc with patient_intake.xlsx
"""

import os
import shlex
import subprocess
import time
import openpyxl
from fpdf import FPDF

WORKDIR = '/home/user'
DESKTOP = '/home/user/Desktop'
TASK_ID = 'osworld_multi_apps_excel_pdf_form_003'
SPREADSHEET = f'{WORKDIR}/patient_intake.xlsx'
TEMPLATE_PDF = f'{DESKTOP}/intake_form_template.pdf'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_spreadsheet():
    """Create patient_intake.xlsx with realistic patient data."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patient Intake"

    # Headers
    headers = ['PatientName', 'DOB', 'InsuranceID', 'PrimaryCondition', 'Assigned Doctor', 'Appointment Date']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic patient data
    patients = [
        ['Maria Garcia',     '1985-03-22', 'INS-447821', 'Hypertension',        'Dr. James Patel',    '2025-06-10'],
        ['David Chen',       '1972-11-08', 'INS-339056', 'Type 2 Diabetes',     'Dr. Emily Torres',   '2025-06-11'],
        ['Sophia Williams',  '1990-07-14', 'INS-552134', 'Asthma',              'Dr. Robert Kim',     '2025-06-12'],
        ['Liam Johnson',     '1965-01-30', 'INS-661789', 'Chronic Back Pain',   'Dr. James Patel',    '2025-06-13'],
        ['Aisha Mohammed',   '1998-09-05', 'INS-774302', 'Anxiety Disorder',    'Dr. Sarah Nguyen',   '2025-06-16'],
    ]

    for r, row_data in enumerate(patients, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Auto-adjust column widths for readability
    for col_idx, col in enumerate(ws.columns, 1):
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

    wb.save(SPREADSHEET)
    print(f'Spreadsheet created: {SPREADSHEET}')


def create_template_pdf():
    """Create a blank PDF intake form template with fillable-looking fields."""
    os.makedirs(DESKTOP, exist_ok=True)

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 12, "Patient Intake Form", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    pdf.set_font("Helvetica", "", 10)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 6, "Please complete all fields accurately.", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    pdf.set_text_color(0, 0, 0)

    def field_line(label, value=""):
        pdf.set_font("Helvetica", "B", 11)
        pdf.cell(60, 8, label + ":", new_x="END", new_y="TOP")
        pdf.set_font("Helvetica", "", 11)
        # Draw underline box for empty field
        x = pdf.get_x()
        y = pdf.get_y()
        pdf.set_draw_color(150, 150, 150)
        pdf.rect(x, y, 120, 8)
        if value:
            pdf.cell(120, 8, value, new_x="LMARGIN", new_y="NEXT")
        else:
            pdf.cell(120, 8, "", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(3)

    field_line("Patient Name")
    field_line("Date of Birth")
    field_line("Insurance ID")
    field_line("Primary Condition")
    field_line("Assigned Doctor")
    field_line("Appointment Date")

    pdf.ln(8)
    pdf.set_font("Helvetica", "I", 9)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, "Clinic Administrative Use Only - All information kept confidential.", new_x="LMARGIN", new_y="NEXT")

    pdf.output(TEMPLATE_PDF)
    print(f'Template PDF created: {TEMPLATE_PDF}')


def main():
    create_spreadsheet()
    create_template_pdf()

    # GUI-ready startup: open LibreOffice Calc with the spreadsheet
    launch_gui(f'libreoffice --calc "{SPREADSHEET}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


main()
