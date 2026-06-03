"""
Reward Script: Color-coded conditional formatting on Grade column
Task ID: calc_gsd_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Conditional formatting exists on E2:E31
  Component 2 (0.25) - Green rule for 'A' grades
  Component 3 (0.25) - Yellow rule for 'B' grades
  Component 4 (0.20) - Red rule for 'F' grades
  Component 5 (0.10) - Grade data unchanged
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_008'

# Expected grade values (ground truth from initial state)
EXPECTED_GRADES = [
    'A', 'B', 'A', 'D', 'C', 'B', 'F', 'A', 'C', 'B',
    'A', 'D', 'B', 'F', 'A', 'C', 'B', 'A', 'D', 'B',
    'C', 'F', 'A', 'B', 'C', 'A', 'D', 'B', 'F', 'B'
]

# Expected conditional formatting colors (8-char ARGB)
EXPECTED_RULES = {
    'A': 'FF70AD47',  # green
    'B': 'FFFFFF00',  # yellow
    'F': 'FFFF0000',  # red
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Grades' sheet must exist
    if 'Grades' not in wb.sheetnames:
        print("FAIL: 'Grades' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Grades']

    # Collect all conditional formatting rules on E2:E31
    cf_rules_on_target = []
    for cf in ws.conditional_formatting:
        range_str = str(cf).replace('<ConditionalFormatting ', '').replace('>', '').strip()
        # Check if E2:E31 is covered by this range
        if 'E2:E31' in range_str or 'E2:E31' in str(cf.sqref):
            for rule in cf.rules:
                cf_rules_on_target.append(rule)

    # Component 1: Conditional formatting rules exist on E2:E31 (0.20 pts)
    try:
        if len(cf_rules_on_target) >= 3:
            print(f"PASS: Component 1 — Found {len(cf_rules_on_target)} CF rules on E2:E31 (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Expected >= 3 CF rules on E2:E31, found {len(cf_rules_on_target)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Helper: find a cellIs/equal rule matching a given grade letter and color
    def find_rule_for_grade(grade_letter, expected_color):
        """Check if there's a CF rule for the given grade with the expected fill color."""
        for rule in cf_rules_on_target:
            rule_type = getattr(rule, 'type', None)
            rule_op = getattr(rule, 'operator', None)
            rule_formula = getattr(rule, 'formula', [])

            # Check if this is a cellIs/equal rule for the right letter
            is_cell_is = rule_type == 'cellIs'
            is_equal = rule_op == 'equal'
            matches_value = any(
                str(f).strip().strip('"').strip("'") == grade_letter
                for f in (rule_formula or [])
            )

            if is_cell_is and is_equal and matches_value:
                # Check fill color
                if rule.dxf and rule.dxf.fill:
                    fg = rule.dxf.fill.fgColor
                    if fg and hasattr(fg, 'rgb') and fg.rgb:
                        actual_color = str(fg.rgb)
                        if actual_color == expected_color:
                            return True, actual_color
                        else:
                            return False, actual_color
                return False, 'no_fill_color'
        return False, 'no_matching_rule'

    # Component 2: Green rule for 'A' grades (0.25 pts)
    try:
        found, detail = find_rule_for_grade('A', EXPECTED_RULES['A'])
        if found:
            print(f"PASS: Component 2 — 'A' grade rule with green fill ({EXPECTED_RULES['A']}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 — 'A' grade rule: {detail}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Yellow rule for 'B' grades (0.25 pts)
    try:
        found, detail = find_rule_for_grade('B', EXPECTED_RULES['B'])
        if found:
            print(f"PASS: Component 3 — 'B' grade rule with yellow fill ({EXPECTED_RULES['B']}) (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 — 'B' grade rule: {detail}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Red rule for 'F' grades (0.20 pts)
    try:
        found, detail = find_rule_for_grade('F', EXPECTED_RULES['F'])
        if found:
            print(f"PASS: Component 4 — 'F' grade rule with red fill ({EXPECTED_RULES['F']}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — 'F' grade rule: {detail}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: CF rules applied AND grade data unchanged (0.10 pts)
    # Compound check: both conditions must hold. Data integrity alone is a precondition,
    # so we gate it on CF rules existing (the task-introduced change).
    try:
        actual_grades = [ws.cell(row=r, column=5).value for r in range(2, 32)]
        data_intact = (actual_grades == EXPECTED_GRADES)
        has_cf = len(cf_rules_on_target) >= 3
        if has_cf and data_intact:
            print(f"PASS: Component 5 — CF rules present AND all 30 grade values intact (0.10 pts)")
            total_score += 0.10
        elif not has_cf:
            print(f"FAIL: Component 5 — No CF rules found, compound check fails")
        else:
            diffs = [(i+2, e, a) for i, (e, a) in enumerate(zip(EXPECTED_GRADES, actual_grades)) if e != a]
            print(f"FAIL: Component 5 — Grade values changed: {diffs[:5]}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 1)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persistence hook: save any unsaved LibreOffice edits before verification
def persist_app_state():
    import time
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


persist_app_state()

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
