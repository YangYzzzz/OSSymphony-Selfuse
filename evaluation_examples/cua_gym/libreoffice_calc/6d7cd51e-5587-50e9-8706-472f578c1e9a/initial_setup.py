"""
Initial Setup: Format Grade column with conditional formatting (color-coded backgrounds)
Task ID: calc_gsd_008
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsd_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Grades"

    # Headers
    headers = ["Student ID", "Name", "Score", "Percentage", "Grade", "Remarks"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 30 student records with realistic data
    students = [
        ["STU-1001", "Sarah Chen", 95, 95.0, "A", "Excellent performance"],
        ["STU-1002", "Marcus Johnson", 82, 82.0, "B", "Good work"],
        ["STU-1003", "Priya Patel", 91, 91.0, "A", "Outstanding"],
        ["STU-1004", "James O'Brien", 67, 67.0, "D", "Needs improvement"],
        ["STU-1005", "Aisha Mohammed", 78, 78.0, "C", "Average"],
        ["STU-1006", "Lucas Fernandez", 88, 88.0, "B", "Very good"],
        ["STU-1007", "Emily Watson", 45, 45.0, "F", "Failed - requires retake"],
        ["STU-1008", "Raj Krishnamurthy", 93, 93.0, "A", "Excellent"],
        ["STU-1009", "Olivia Brown", 71, 71.0, "C", "Satisfactory"],
        ["STU-1010", "David Kim", 86, 86.0, "B", "Good progress"],
        ["STU-1011", "Sofia Martinez", 98, 98.0, "A", "Top of class"],
        ["STU-1012", "William Taylor", 55, 55.0, "D", "Below average"],
        ["STU-1013", "Fatima Al-Hassan", 84, 84.0, "B", "Consistent effort"],
        ["STU-1014", "Noah Anderson", 39, 39.0, "F", "Failed - see counselor"],
        ["STU-1015", "Yuki Tanaka", 92, 92.0, "A", "Excellent analytical skills"],
        ["STU-1016", "Grace Okonkwo", 76, 76.0, "C", "Improving steadily"],
        ["STU-1017", "Liam Murphy", 81, 81.0, "B", "Good understanding"],
        ["STU-1018", "Mei-Lin Zhou", 97, 97.0, "A", "Exceptional work"],
        ["STU-1019", "Carlos Rivera", 63, 63.0, "D", "Needs more practice"],
        ["STU-1020", "Hannah Schmidt", 89, 89.0, "B", "Very strong"],
        ["STU-1021", "Alexander Petrov", 73, 73.0, "C", "Moderate performance"],
        ["STU-1022", "Isabella Garcia", 42, 42.0, "F", "Failed - retake scheduled"],
        ["STU-1023", "Ethan Wright", 90, 90.0, "A", "Outstanding effort"],
        ["STU-1024", "Zara Hussain", 85, 85.0, "B", "Well done"],
        ["STU-1025", "Daniel Larsson", 77, 77.0, "C", "Fair performance"],
        ["STU-1026", "Amara Diallo", 94, 94.0, "A", "Remarkable achievement"],
        ["STU-1027", "Ryan Cooper", 61, 61.0, "D", "Struggling with material"],
        ["STU-1028", "Naomi Sato", 83, 83.0, "B", "Steady improvement"],
        ["STU-1029", "Benjamin Lee", 48, 48.0, "F", "Failed - needs support"],
        ["STU-1030", "Chloe Dubois", 87, 87.0, "B", "Great participation"],
    ]

    for r, row_data in enumerate(students, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30

    # NO conditional formatting - that is the task for the agent
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
