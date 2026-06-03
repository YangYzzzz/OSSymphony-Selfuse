"""
Reward Script: Segment customer base by annual spend into tiers using IFS formula,
apply color coding, create SUMIFS aggregation, and add pie chart.
Task ID: calc_sales_customer_segment_021
Domain: libreoffice_calc
Scoring:
  - Component 1: IFS formula in E2:E201 (0.30 pts)
  - Component 2: Conditional formatting on E2:E201 with tier colors (0.30 pts)
  - Component 3: SUMIFS formulas in TierSummary B2:B5 (0.20 pts)
  - Component 4: Pie chart on TierSummary with correct title (0.20 pts)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_customer_segment_021'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook - precondition gate
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check required sheets exist - precondition gate
    if 'Customers' not in wb.sheetnames or 'TierSummary' not in wb.sheetnames:
        print(f"CRITICAL: Required sheets not found. Found: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws_customers = wb['Customers']
    ws_summary = wb['TierSummary']

    # Component 1: IFS formulas in Customers!E2:E201 (0.30 points)
    # The task requires IFS formulas assigning tier based on D column values
    # Thresholds: Platinum (D>=500000), Gold (D>=200000), Silver (D>=50000), Bronze (else)
    # This FAILS on initial (E column is empty) and PASSES on golden
    try:
        ifs_count = 0
        correct_ifs_count = 0
        for row in range(2, 202):  # rows 2-201 for 200 customers
            cell_val = ws_customers.cell(row=row, column=5).value  # Column E
            if cell_val is not None:
                ifs_count += 1
                val_str = str(cell_val).upper().replace(' ', '')
                # Check that it's an IFS formula referencing D column thresholds
                if (val_str.startswith('=IFS') and
                        '500000' in val_str and
                        '200000' in val_str and
                        '50000' in val_str and
                        'PLATINUM' in val_str and
                        'GOLD' in val_str and
                        'SILVER' in val_str and
                        'BRONZE' in val_str):
                    correct_ifs_count += 1

        if correct_ifs_count == 200:
            print(f"PASS: Component 1 — IFS formula in all 200 E cells with correct tier thresholds (0.30 pts)")
            total_score += 0.30
        elif correct_ifs_count >= 150:
            print(f"PARTIAL: Component 1 — IFS formula in {correct_ifs_count}/200 cells (partial 0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — IFS formula found in only {correct_ifs_count}/200 cells (expected 200); non-IFS cells: {200 - ifs_count}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Conditional formatting on E2:E201 with tier-specific colors (0.30 points)
    # Task requires: platinum=purple, gold=gold, silver=gray, bronze=brown fill
    # This FAILS on initial (no CF rules) and PASSES on golden
    try:
        cf_rules = ws_customers.conditional_formatting._cf_rules
        # Look for a CF range covering E2:E201 (or equivalent)
        tier_rules_found = {
            'Platinum': False,
            'Gold': False,
            'Silver': False,
            'Bronze': False,
        }

        for cf_range, rules_list in cf_rules.items():
            # Check that range covers E column cells
            range_str = str(cf_range)
            if 'E' not in range_str:
                continue
            for rule in rules_list:
                formula_list = getattr(rule, 'formula', []) or []
                formula_str = ' '.join(str(f) for f in formula_list).upper()
                try:
                    has_fill = (rule.dxf is not None and
                                rule.dxf.fill is not None and
                                rule.dxf.fill.fgColor is not None)
                except Exception:
                    has_fill = False

                if has_fill and formula_list:
                    if 'PLATINUM' in formula_str:
                        tier_rules_found['Platinum'] = True
                    if 'GOLD' in formula_str:
                        tier_rules_found['Gold'] = True
                    if 'SILVER' in formula_str:
                        tier_rules_found['Silver'] = True
                    if 'BRONZE' in formula_str:
                        tier_rules_found['Bronze'] = True

        tiers_with_cf = sum(1 for v in tier_rules_found.values() if v)

        if tiers_with_cf == 4:
            print(f"PASS: Component 2 — Conditional formatting found for all 4 tiers (0.30 pts)")
            total_score += 0.30
        elif tiers_with_cf >= 2:
            partial = round(0.30 * tiers_with_cf / 4, 2)
            print(f"PARTIAL: Component 2 — CF found for {tiers_with_cf}/4 tiers: {[k for k,v in tier_rules_found.items() if v]} ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — CF found for only {tiers_with_cf}/4 tiers: {tier_rules_found}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: SUMIFS formulas in TierSummary B2:B5 (0.20 points)
    # Task requires aggregating Annual Spend by tier in TierSummary sheet
    # This FAILS on initial (B2:B5 is empty) and PASSES on golden
    try:
        tier_names = ['Platinum', 'Gold', 'Silver', 'Bronze']
        sumifs_count = 0
        for row_idx, tier_name in enumerate(tier_names, start=2):
            b_cell = ws_summary.cell(row=row_idx, column=2).value
            if b_cell is not None:
                b_str = str(b_cell).upper().replace(' ', '')
                if ('SUMIFS' in b_str and
                        'CUSTOMERS' in b_str and
                        tier_name.upper() in b_str):
                    sumifs_count += 1
                    print(f"  PASS: B{row_idx} has SUMIFS for {tier_name}: {b_cell}")
                else:
                    print(f"  FAIL: B{row_idx} expected SUMIFS for {tier_name}, found: {b_cell}")
            else:
                print(f"  FAIL: B{row_idx} is empty (expected SUMIFS for {tier_name})")

        if sumifs_count == 4:
            print(f"PASS: Component 3 — All 4 SUMIFS formulas present in TierSummary B2:B5 (0.20 pts)")
            total_score += 0.20
        elif sumifs_count >= 2:
            partial = round(0.20 * sumifs_count / 4, 2)
            print(f"PARTIAL: Component 3 — {sumifs_count}/4 SUMIFS formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Only {sumifs_count}/4 SUMIFS formulas found")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Pie chart on TierSummary with title 'Revenue by Customer Tier' (0.20 points)
    # Task explicitly requires a pie chart showing revenue share per tier with this title
    # This FAILS on initial (no chart) and PASSES on golden
    try:
        charts = ws_summary._charts
        pie_charts = [c for c in charts if type(c).__name__ == 'PieChart']

        if not pie_charts:
            print(f"FAIL: Component 4 — No pie chart found on TierSummary (found {len(charts)} total charts)")
        else:
            pie = pie_charts[0]
            # Extract chart title text
            chart_title_text = ''
            try:
                chart_title_text = pie.title.tx.rich.p[0].r[0].t
            except Exception:
                try:
                    # Fallback: search str representation for the title string
                    title_repr = str(pie.title)
                    expected = 'Revenue by Customer Tier'
                    if expected in title_repr:
                        chart_title_text = expected
                except Exception:
                    chart_title_text = ''

            expected_title = 'Revenue by Customer Tier'
            title_matches = expected_title.lower() in chart_title_text.lower()
            if title_matches:
                print(f"  PASS: Pie chart title matches: '{chart_title_text}'")
                print(f"PASS: Component 4 — Pie chart with correct title found (0.20 pts)")
                total_score += 0.20
            elif len(pie_charts) > 0:
                total_score += 0.10
                print(f"  FAIL: Pie chart title is '{chart_title_text}', expected '{expected_title}'")
                print(f"PARTIAL: Component 4 — Pie chart exists but title mismatch (0.10 pts)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {round(total_score, 2)}/1.0")
    print(f"REWARD: {round(final_score, 2)}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
