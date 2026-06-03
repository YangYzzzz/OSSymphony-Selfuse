"""
Initial Setup: Customer segmentation spreadsheet with annual spend data
Task ID: calc_sales_customer_segment_021
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_customer_segment_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Customers ---
    ws1 = wb.active
    ws1.title = 'Customers'

    # Headers
    headers = ['Customer ID', 'Company', 'Industry', 'Annual Spend', 'Tier', 'Rep Assigned']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)

    # 200 rows of realistic customer data
    # Annual spend distributed across tiers:
    # Platinum ($500K+): ~20 customers
    # Gold ($200K-$499K): ~40 customers
    # Silver ($50K-$199K): ~80 customers
    # Bronze (under $50K): ~60 customers

    companies = [
        ('Apex Dynamics', 'Technology'), ('Blue Ridge Partners', 'Finance'),
        ('Cornerstone Medical', 'Healthcare'), ('Delta Logistics', 'Transportation'),
        ('Eagle Eye Security', 'Technology'), ('Frontier Analytics', 'Consulting'),
        ('GreenPath Energy', 'Energy'), ('Harbor View Retail', 'Retail'),
        ('Integra Systems', 'Technology'), ('Jasper Capital', 'Finance'),
        ('Keystone Manufacturing', 'Manufacturing'), ('Lighthouse Media', 'Media'),
        ('Meridian Healthcare', 'Healthcare'), ('Northgate Construction', 'Construction'),
        ('Oceanic Ventures', 'Finance'), ('Pinnacle Software', 'Technology'),
        ('Quantum Research', 'Research'), ('Riverside Foods', 'Food & Beverage'),
        ('Summit Consulting', 'Consulting'), ('TechBridge Solutions', 'Technology'),
        ('United Distribution', 'Distribution'), ('Vanguard Industries', 'Manufacturing'),
        ('Westfield Properties', 'Real Estate'), ('Xcell Pharmaceuticals', 'Healthcare'),
        ('Zenith Aerospace', 'Aerospace'), ('Alliant Financial', 'Finance'),
        ('Beacon Education', 'Education'), ('Cascade Networks', 'Technology'),
        ('Diamond Retail Group', 'Retail'), ('Eclipse Engineering', 'Engineering'),
        ('Falcon Security', 'Security'), ('Global Trade Corp', 'Trade'),
        ('Highland Brewing', 'Food & Beverage'), ('Irongate Steel', 'Manufacturing'),
        ('Jade Cosmetics', 'Retail'), ('Kingston Legal', 'Legal'),
        ('Lunar Labs', 'Research'), ('Maple Grove Hotels', 'Hospitality'),
        ('Nexus Communications', 'Telecommunications'), ('Omega Financial', 'Finance'),
        ('Pacific Rim Imports', 'Trade'), ('Quest Diagnostics Group', 'Healthcare'),
        ('Redwood Architecture', 'Architecture'), ('Silverline Transit', 'Transportation'),
        ('Titan Energy Solutions', 'Energy'), ('Urban Design Studio', 'Design'),
        ('Valor Defense Systems', 'Aerospace'), ('Wavelength Media', 'Media'),
        ('X-Caliber Tools', 'Manufacturing'), ('Yellowstone Tourism', 'Hospitality'),
    ]

    reps = [
        'Amanda Torres', 'Brian Mitchell', 'Carlos Ruiz', 'Diana Chen',
        'Eric Nakamura', 'Fiona Walsh', 'George Patel', 'Hannah Kim',
        'Ivan Kozlov', 'Julia Santos'
    ]

    # Generate 200 customers with varied spend amounts
    import random
    random.seed(42)

    # Spend amounts: mix of tiers
    spend_ranges = (
        # Platinum (500K-1.2M): 20 customers
        [(random.randint(500000, 1200000)) for _ in range(20)] +
        # Gold (200K-499K): 40 customers
        [(random.randint(200000, 499000)) for _ in range(40)] +
        # Silver (50K-199K): 80 customers
        [(random.randint(50000, 199000)) for _ in range(80)] +
        # Bronze (8K-49K): 60 customers
        [(random.randint(8000, 49000)) for _ in range(60)]
    )
    random.shuffle(spend_ranges)

    industries = [
        'Technology', 'Finance', 'Healthcare', 'Manufacturing', 'Retail',
        'Consulting', 'Energy', 'Transportation', 'Real Estate', 'Education',
        'Media', 'Construction', 'Food & Beverage', 'Aerospace', 'Legal'
    ]

    for i in range(200):
        row = i + 2
        cust_id = f'CUST-{1001 + i}'
        comp_idx = i % len(companies)
        company = companies[comp_idx][0]
        industry = industries[i % len(industries)]
        spend = spend_ranges[i]
        tier = ''  # Empty - task asks to fill this
        rep = reps[i % len(reps)]

        ws1.cell(row=row, column=1, value=cust_id)
        ws1.cell(row=row, column=2, value=company)
        ws1.cell(row=row, column=3, value=industry)
        ws1.cell(row=row, column=4, value=spend)
        ws1.cell(row=row, column=5, value=tier)  # Empty
        ws1.cell(row=row, column=6, value=rep)

    # Format Annual Spend column as currency
    for row in range(2, 202):
        ws1.cell(row=row, column=4).number_format = '$#,##0'

    # Column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 28
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 16
    ws1.column_dimensions['E'].width = 12
    ws1.column_dimensions['F'].width = 20

    # Freeze row 1
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: TierSummary ---
    ws2 = wb.create_sheet('TierSummary')

    # Headers
    ws2['A1'] = 'Tier'
    ws2['B1'] = 'Total Revenue'
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)

    # Tier labels
    tiers = ['Platinum', 'Gold', 'Silver', 'Bronze']
    for i, tier in enumerate(tiers):
        ws2.cell(row=i + 2, column=1, value=tier)

    # B2:B5 left empty — task asks to add SUMIFS here
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Customers sheet: 200 rows, 6 columns')
    print(f'  TierSummary sheet: tier labels in A2:A5, B2:B5 empty')
    print(f'  Column E (Tier) is empty - ready for IFS formula')

create_initial()
