"""
Initial Setup: Conference Registration Tracker
Task ID: calc_wf_073
Domain: libreoffice_calc

Creates a spreadsheet with 65 conference registrations across 4 sessions.
No formulas, charts, conditional formatting, or badge area — those are the task.
"""

import os
import shlex
import subprocess
import time
import random
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_073'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registrations"

    # --- Headers ---
    headers = [
        "Name", "Email", "Company", "Session Choice",
        "Dietary", "Payment", "Registration Date", "Amount"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Realistic Data ---
    first_names = [
        "Sarah", "Marcus", "Li", "Priya", "James", "Yuki", "Elena", "Omar",
        "Chen", "Aisha", "David", "Maria", "Raj", "Fatima", "Thomas",
        "Sofia", "Kenji", "Amara", "Lucas", "Nina", "Wei", "Isabella",
        "Ahmed", "Megan", "Carlos", "Hannah", "Dmitri", "Zara", "Patrick",
        "Leila", "Robert", "Ananya", "Felix", "Grace", "Hiroshi",
        "Valentina", "Samuel", "Ingrid", "Mohammad", "Charlotte",
        "Takeshi", "Emma", "Olga", "Daniel", "Sana", "Vincent",
        "Beatriz", "Nathan", "Chloe", "Arjun", "Lena", "Gabriel",
        "Yara", "Kevin", "Simone", "Tariq", "Julia", "Ivan",
        "Nadia", "Eric", "Mei", "Alexander", "Rosa", "Finn", "Layla"
    ]

    last_names = [
        "Chen", "Johnson", "Wei", "Patel", "Morrison", "Tanaka", "Volkov",
        "Hassan", "Liu", "Okafor", "Brennan", "Garcia", "Sharma", "Al-Rashid",
        "Klein", "Rossi", "Nakamura", "Diallo", "Fernandez", "Bergstrom",
        "Zhang", "Torres", "Khoury", "Walsh", "Reyes", "Mueller", "Petrov",
        "Osei", "O'Brien", "Mansouri", "Park", "Kapoor", "Lindgren", "Obi",
        "Sato", "Popov", "Brooks", "Ericsson", "Rahman", "Dumont",
        "Watanabe", "Bennett", "Sokolova", "Foster", "Abbasi", "Lambert",
        "Oliveira", "Carlson", "Mercier", "Rao", "Fischer", "Almeida",
        "Nazari", "Wu", "Dubois", "Malik", "Hoffman", "Volkov",
        "Johansson", "Harper", "Huang", "Schmidt", "Ferreira", "O'Connor", "Ibrahim"
    ]

    companies = [
        "TechVista Solutions", "Quantum Analytics", "DataBridge Corp",
        "NexGen Systems", "CloudPeak Inc", "SynergyTech", "Apex Digital",
        "InnoWave Labs", "CoreStack Technologies", "BrightPath AI",
        "Meridian Consulting", "PulsePoint Media", "Vertex Dynamics",
        "BlueSky Research", "Catalyst Ventures", "Horizon Partners",
        "Luminary Group", "Prism Analytics", "Forge Innovations",
        "Atlas Networks", "Summit Software", "Pinnacle Data",
        "Mosaic Digital", "Vanguard Systems", "Ember Technologies"
    ]

    sessions = ["A", "B", "C", "D"]
    # Weight sessions to make some full/near capacity
    # A=50 cap, B=30 cap, C=40 cap, D=35 cap
    session_weights = [0.30, 0.28, 0.22, 0.20]

    dietary_options = ["None", "Vegetarian", "Vegan", "Gluten-Free"]
    dietary_weights = [0.55, 0.20, 0.12, 0.13]

    payment_options = ["Paid", "Unpaid", "Partial"]
    payment_weights = [0.60, 0.25, 0.15]

    amounts = {
        "Paid": [450.00, 500.00, 550.00, 475.00, 525.00],
        "Unpaid": [0.00],
        "Partial": [150.00, 200.00, 225.00, 250.00, 175.00],
    }

    # Generate 65 registrations over a date range
    base_date = datetime(2026, 2, 1)
    end_date = datetime(2026, 3, 25)
    date_range_days = (end_date - base_date).days

    data = []
    for i in range(65):
        name = f"{first_names[i]} {last_names[i]}"
        email_first = first_names[i].lower()
        email_last = last_names[i].lower().replace("'", "")
        company = random.choice(companies)
        email = f"{email_first}.{email_last}@{company.lower().replace(' ', '').replace(',', '')}.com"

        session = random.choices(sessions, weights=session_weights, k=1)[0]
        dietary = random.choices(dietary_options, weights=dietary_weights, k=1)[0]
        payment = random.choices(payment_options, weights=payment_weights, k=1)[0]

        # Spread dates across range with slight clustering toward later dates
        day_offset = int(random.triangular(0, date_range_days, date_range_days * 0.7))
        reg_date = base_date + timedelta(days=day_offset)

        amount = random.choice(amounts[payment])

        data.append([name, email, company, session, dietary, payment, reg_date, amount])

    # Sort by registration date for realism
    data.sort(key=lambda x: x[6])

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border
            if c == 7:  # Date column
                cell.number_format = 'yyyy-mm-dd'
            elif c == 8:  # Amount column
                cell.number_format = '$#,##0.00'
            if c in (4, 5, 6):  # Center-align categorical
                cell.alignment = Alignment(horizontal="center")

    # Set column widths
    col_widths = {"A": 22, "B": 38, "C": 24, "D": 15, "E": 14, "F": 12, "G": 16, "H": 12}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
