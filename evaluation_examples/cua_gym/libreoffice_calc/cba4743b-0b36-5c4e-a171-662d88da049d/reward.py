"""
Reward Script: Conference Registration Tracker
Task ID: calc_wf_073
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): Status/Waitlist formulas in column I
  Component 2 (0.20): Session Summary with COUNTIF formulas (K1:N7)
  Component 3 (0.15): Payment Summary with SUMIF/COUNTIF formulas (K9:M14)
  Component 4 (0.10): Dietary Summary with COUNTIF formulas (K16:L22)
  Component 5 (0.15): Conditional formatting (yellow Unpaid, red Waitlist)
  Component 6 (0.10): Line chart for cumulative registrations
  Component 7 (0.10): Badges sheet with name badge layout
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_073'


def persist_app_state(domain):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Registrations' sheet must exist
    if 'Registrations' not in wb.sheetnames:
        print("CRITICAL: 'Registrations' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Registrations']

    # Component 1: Status/Waitlist formulas in column I (0.20 points)
    # The golden file has IF(COUNTIF(...)) waitlist formulas in column I for rows 2-66
    try:
        formula_count = 0
        has_waitlist_logic = False
        for r in range(2, 67):
            val = ws.cell(row=r, column=9).value
            if val is not None and isinstance(val, str) and val.startswith('='):
                formula_count += 1
                upper_val = val.upper()
                if 'COUNTIF' in upper_val and 'WAITLIST' in upper_val:
                    has_waitlist_logic = True

        if formula_count >= 60 and has_waitlist_logic:
            print(f"PASS: Component 1 -- Status formulas found ({formula_count}/65 rows with waitlist logic) (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 30 and has_waitlist_logic:
            print(f"PARTIAL: Component 1 -- Status formulas partially present ({formula_count}/65) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 -- Expected waitlist formulas in col I, found {formula_count} formulas, waitlist_logic={has_waitlist_logic}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Session Summary with COUNTIF formulas (0.20 points)
    # Golden: K1='Session Summary', K2='Session', L2='Count', M2='Capacity'
    # K3-K6 = A/B/C/D, L3-L6 = COUNTIF formulas, M3-M6 = capacity numbers
    try:
        has_session_header = False
        has_countif_formulas = False
        has_capacity_values = False

        # Check for session summary header
        for r in range(1, 10):
            val = ws.cell(row=r, column=11).value
            if val is not None and 'session' in str(val).lower() and 'summar' in str(val).lower():
                has_session_header = True
                break

        # Check for COUNTIF formulas in session count area (L column near session labels)
        countif_count = 0
        for r in range(1, 25):
            val = ws.cell(row=r, column=12).value
            if val is not None and isinstance(val, str) and 'COUNTIF' in val.upper():
                countif_count += 1

        if countif_count >= 4:
            has_countif_formulas = True

        # Check for capacity values (50, 30, 40, 35)
        expected_caps = {50, 30, 40, 35}
        found_caps = set()
        for r in range(1, 25):
            val = ws.cell(row=r, column=13).value
            if val is not None:
                try:
                    v = int(val)
                    if v in expected_caps:
                        found_caps.add(v)
                except (ValueError, TypeError):
                    pass
        if len(found_caps) >= 3:
            has_capacity_values = True

        if has_session_header and has_countif_formulas and has_capacity_values:
            print(f"PASS: Component 2 -- Session Summary complete (header, {countif_count} COUNTIFs, capacities {found_caps}) (0.20 pts)")
            total_score += 0.20
        elif has_countif_formulas:
            print(f"PARTIAL: Component 2 -- Session COUNTIF formulas found but incomplete (header={has_session_header}, caps={found_caps}) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 -- Session Summary missing (header={has_session_header}, countifs={countif_count}, caps={found_caps})")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Payment Summary with SUMIF/COUNTIF formulas (0.15 points)
    # Golden: K9='Payment Summary', K11='Paid', K12='Unpaid', K13='Partial'
    # L11-L13 = COUNTIF, M11-M13 = SUMIF
    try:
        has_payment_header = False
        has_payment_countif = False
        has_payment_sumif = False

        for r in range(1, 25):
            val = ws.cell(row=r, column=11).value
            if val is not None and 'payment' in str(val).lower() and 'summar' in str(val).lower():
                has_payment_header = True
                break

        # Look for SUMIF formulas in column M
        sumif_count = 0
        payment_countif_count = 0
        for r in range(1, 25):
            m_val = ws.cell(row=r, column=13).value
            l_val = ws.cell(row=r, column=12).value
            if m_val is not None and isinstance(m_val, str) and 'SUMIF' in m_val.upper():
                sumif_count += 1
            if l_val is not None and isinstance(l_val, str) and 'COUNTIF' in l_val.upper():
                # Check if this is in the payment section (near Paid/Unpaid/Partial labels)
                k_val = ws.cell(row=r, column=11).value
                if k_val is not None and str(k_val).strip().lower() in ('paid', 'unpaid', 'partial'):
                    payment_countif_count += 1

        if sumif_count >= 2:
            has_payment_sumif = True
        if payment_countif_count >= 2:
            has_payment_countif = True

        if has_payment_header and has_payment_sumif and has_payment_countif:
            print(f"PASS: Component 3 -- Payment Summary complete ({sumif_count} SUMIFs, {payment_countif_count} COUNTIFs) (0.15 pts)")
            total_score += 0.15
        elif has_payment_sumif or has_payment_countif:
            print(f"PARTIAL: Component 3 -- Payment Summary partial (header={has_payment_header}, sumifs={sumif_count}, countifs={payment_countif_count}) (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 3 -- Payment Summary missing")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Dietary Summary with COUNTIF formulas (0.10 points)
    # Golden: K16='Dietary Summary', K18='None', K19='Vegetarian', K20='Vegan', K21='Gluten-Free'
    try:
        has_dietary_header = False
        dietary_countif_count = 0

        for r in range(1, 30):
            val = ws.cell(row=r, column=11).value
            if val is not None and 'dietary' in str(val).lower() and 'summar' in str(val).lower():
                has_dietary_header = True
                break

        # Count COUNTIF formulas associated with dietary labels
        dietary_types = {'none', 'vegetarian', 'vegan', 'gluten-free', 'gf', 'gluten free'}
        for r in range(1, 30):
            l_val = ws.cell(row=r, column=12).value
            k_val = ws.cell(row=r, column=11).value
            if l_val is not None and isinstance(l_val, str) and 'COUNTIF' in l_val.upper():
                if k_val is not None and str(k_val).strip().lower() in dietary_types:
                    dietary_countif_count += 1

        if has_dietary_header and dietary_countif_count >= 3:
            print(f"PASS: Component 4 -- Dietary Summary complete ({dietary_countif_count} COUNTIFs) (0.10 pts)")
            total_score += 0.10
        elif dietary_countif_count >= 2:
            print(f"PARTIAL: Component 4 -- Dietary COUNTIFs found but incomplete (header={has_dietary_header}, count={dietary_countif_count}) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 -- Dietary Summary missing (header={has_dietary_header}, countifs={dietary_countif_count})")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Conditional formatting (0.15 points)
    # Golden: expression rules for $F2="Unpaid" (yellow) and $I2="Waitlist" (red)
    try:
        has_unpaid_cf = False
        has_waitlist_cf = False

        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                formula_str = str(rule.formula).upper() if rule.formula else ''
                if 'UNPAID' in formula_str:
                    has_unpaid_cf = True
                if 'WAITLIST' in formula_str:
                    has_waitlist_cf = True

        if has_unpaid_cf and has_waitlist_cf:
            print(f"PASS: Component 5 -- Both conditional formatting rules found (Unpaid + Waitlist) (0.15 pts)")
            total_score += 0.15
        elif has_unpaid_cf or has_waitlist_cf:
            print(f"PARTIAL: Component 5 -- Only one CF rule found (Unpaid={has_unpaid_cf}, Waitlist={has_waitlist_cf}) (0.07 pts)")
            total_score += 0.07
        else:
            print(f"FAIL: Component 5 -- No conditional formatting for Unpaid/Waitlist found")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    # Component 6: Line chart for cumulative registrations (0.10 points)
    # Golden: LineChart on Registrations sheet with title containing 'Cumulative Registrations'
    try:
        has_line_chart = False
        chart_on_any_sheet = False

        # Check all sheets for charts
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for ch in sheet._charts:
                chart_on_any_sheet = True
                ctype = type(ch).__name__
                if 'Line' in ctype:
                    has_line_chart = True
                    break
            if has_line_chart:
                break

        if has_line_chart:
            print(f"PASS: Component 6 -- Line chart found (0.10 pts)")
            total_score += 0.10
        elif chart_on_any_sheet:
            # Some chart exists, just not a line chart
            print(f"PARTIAL: Component 6 -- Chart found but not a line chart (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 -- No chart found in any sheet")
    except Exception as e:
        print(f"ERROR: Component 6 -- {e}")

    # Component 7: Badges sheet with name badge layout (0.10 points)
    # Golden: 'Badges' sheet with badge header, fields referencing Registrations sheet
    try:
        has_badges_sheet = False
        has_badge_references = False

        for sname in wb.sheetnames:
            if 'badge' in sname.lower():
                has_badges_sheet = True
                bws = wb[sname]
                # Check for formulas referencing Registrations sheet
                ref_count = 0
                for r in range(1, min(bws.max_row + 1, 75)):
                    for c in range(1, min(bws.max_column + 1, 6)):
                        val = bws.cell(row=r, column=c).value
                        if val is not None and isinstance(val, str) and 'Registrations!' in val:
                            ref_count += 1
                if ref_count >= 10:
                    has_badge_references = True
                print(f"  Badge sheet '{sname}' has {ref_count} references to Registrations")
                break

        if has_badges_sheet and has_badge_references:
            print(f"PASS: Component 7 -- Badges sheet with registration references found (0.10 pts)")
            total_score += 0.10
        elif has_badges_sheet:
            print(f"PARTIAL: Component 7 -- Badges sheet exists but few registration references (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 -- No Badges sheet found")
    except Exception as e:
        print(f"ERROR: Component 7 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
