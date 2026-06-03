"""
Initial Setup: Resource allocation matrix - pre-task state
Task ID: calc_gpm_015
Domain: libreoffice_calc

Creates a Resources sheet with team member data, headers, and capacity values.
Does NOT include: SUM formulas, utilization formulas, conditional formatting,
thick borders, or project totals row.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_015'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resources"

    # --- Merged title row ---
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Resource Allocation - Q2 2026"
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill(start_color="FF006400", end_color="FF006400", fill_type="solid")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    # --- Row 3: Headers ---
    headers = ["Team Member", "Project A", "Project B", "Project C",
               "Project D", "Total Hours", "Capacity", "Utilization"]
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill(start_color="FF90EE90", end_color="FF90EE90", fill_type="solid")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = header_fill
        cell.border = header_border

    # --- Data rows 4-8 ---
    data = [
        ["Sarah", 15, 10, 20, 0,  None, 40, None],
        ["Mike",  20, 15,  0, 12, None, 40, None],
        ["Lisa",   8, 20, 10,  8, None, 40, None],
        ["Tom",    0, 25, 15, 10, None, 40, None],
        ["Amy",   12,  0, 25, 15, None, 40, None],
    ]

    for r_offset, row_data in enumerate(data):
        row_num = 4 + r_offset
        for c_idx, val in enumerate(row_data, 1):
            if val is not None:
                ws.cell(row=row_num, column=c_idx, value=val)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 16
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
