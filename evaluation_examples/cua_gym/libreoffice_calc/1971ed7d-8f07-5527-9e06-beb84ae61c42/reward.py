"""
Reward Script: Resource Allocation Matrix with Conditional Formatting
Task ID: calc_gpm_015
Domain: libreoffice_calc
Scoring:
  Component 1: SUM formulas in F4:F8 (Total Hours)          — 0.25 pts
  Component 2: Utilization formulas in H4:H8 with % format  — 0.20 pts
  Component 3: Conditional formatting on H4:H8 (red/orange/green) — 0.15 pts
  Component 4: Color scale conditional formatting on B4:E8   — 0.10 pts
  Component 5: Project Totals row 10 (A10 label + B10:E10 SUM formulas, bold, top border) — 0.20 pts
  Component 6: Thick borders around matrix                   — 0.10 pts
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_015'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Resources' sheet must exist
    if 'Resources' not in wb.sheetnames:
        print("FAIL: 'Resources' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Resources']

    # Component 1: SUM formulas in F4:F8 (0.25 points)
    # These cells should contain =SUM(B<row>:E<row>) formulas
    try:
        sum_formula_count = 0
        for r in range(4, 9):
            val = ws.cell(row=r, column=6).value  # Column F
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                expected = f"=SUM(B{r}:E{r})"
                if normalized == expected:
                    sum_formula_count += 1
                else:
                    print(f"  F{r}: found '{val}', expected '{expected}'")
            else:
                print(f"  F{r}: no formula found (value: {repr(val)})")

        if sum_formula_count == 5:
            print(f"PASS: Component 1 — All 5 SUM formulas in F4:F8 (0.25 pts)")
            total_score += 0.25
        elif sum_formula_count >= 3:
            partial = round(0.25 * sum_formula_count / 5, 2)
            print(f"PARTIAL: Component 1 — {sum_formula_count}/5 SUM formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {sum_formula_count}/5 SUM formulas in F4:F8")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Utilization formulas in H4:H8 with percentage format (0.20 points)
    try:
        util_formula_count = 0
        pct_format_count = 0
        for r in range(4, 9):
            cell = ws.cell(row=r, column=8)  # Column H
            val = cell.value
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                expected = f"=F{r}/G{r}"
                if normalized == expected:
                    util_formula_count += 1
                else:
                    print(f"  H{r}: found '{val}', expected '{expected}'")
            else:
                print(f"  H{r}: no formula found (value: {repr(val)})")

            # Check percentage number format
            nf = cell.number_format
            if nf and '%' in str(nf):
                pct_format_count += 1

        formula_pts = 0.0
        if util_formula_count == 5:
            formula_pts = 0.12
            print(f"PASS: Component 2a — All 5 utilization formulas in H4:H8")
        elif util_formula_count >= 3:
            formula_pts = round(0.12 * util_formula_count / 5, 2)
            print(f"PARTIAL: Component 2a — {util_formula_count}/5 utilization formulas")
        else:
            print(f"FAIL: Component 2a — Only {util_formula_count}/5 utilization formulas")

        format_pts = 0.0
        if pct_format_count == 5:
            format_pts = 0.08
            print(f"PASS: Component 2b — All 5 H cells have percentage format")
        elif pct_format_count >= 3:
            format_pts = round(0.08 * pct_format_count / 5, 2)
            print(f"PARTIAL: Component 2b — {pct_format_count}/5 cells have percentage format")
        else:
            print(f"FAIL: Component 2b — Only {pct_format_count}/5 cells have percentage format")

        total_score += formula_pts + format_pts
        print(f"  Component 2 total: {formula_pts + format_pts:.2f}/0.20 pts")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Conditional formatting on H4:H8 — red/orange/green (0.15 points)
    try:
        found_h_cf = False
        h_rule_count = 0
        has_greater_than_1 = False
        has_between_09_1 = False
        has_less_than_09 = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if this CF applies to H4:H8
            if 'H4' in cf_range and 'H8' in cf_range:
                found_h_cf = True
                for rule in cf.rules:
                    if rule.type == 'cellIs':
                        h_rule_count += 1
                        op = getattr(rule, 'operator', None)
                        formula = getattr(rule, 'formula', [])
                        if op == 'greaterThan' and formula and str(formula[0]) == '1':
                            has_greater_than_1 = True
                        elif op == 'between' and formula and len(formula) >= 2:
                            if str(formula[0]) == '0.9' and str(formula[1]) == '1':
                                has_between_09_1 = True
                        elif op == 'lessThan' and formula and str(formula[0]) == '0.9':
                            has_less_than_09 = True

        if found_h_cf and has_greater_than_1 and has_between_09_1 and has_less_than_09:
            print(f"PASS: Component 3 — All 3 conditional formatting rules on H4:H8 (0.15 pts)")
            total_score += 0.15
        elif found_h_cf and h_rule_count >= 1:
            matched = sum([has_greater_than_1, has_between_09_1, has_less_than_09])
            partial = round(0.15 * matched / 3, 2)
            print(f"PARTIAL: Component 3 — {matched}/3 CF rules on H4:H8 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No conditional formatting found on H4:H8")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Color scale conditional formatting on B4:E8 (0.10 points)
    try:
        found_color_scale = False
        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            if 'B4' in cf_range and 'E8' in cf_range:
                for rule in cf.rules:
                    if rule.type == 'colorScale':
                        found_color_scale = True
                        break
            if found_color_scale:
                break

        if found_color_scale:
            print(f"PASS: Component 4 — Color scale conditional formatting on B4:E8 (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 4 — No color scale conditional formatting found on B4:E8")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Project Totals row 10 (0.20 points)
    # A10='Project Totals' bold, B10:E10 have SUM formulas, bold, top border
    try:
        comp5_score = 0.0

        # Check A10 label
        a10_val = ws['A10'].value
        a10_bold = ws['A10'].font.bold
        if a10_val and 'Project Totals' in str(a10_val) and a10_bold:
            comp5_score += 0.05
            print(f"  PASS: A10 = '{a10_val}' bold")
        else:
            print(f"  FAIL: A10 expected 'Project Totals' bold, found '{a10_val}' bold={a10_bold}")

        # Check B10:E10 SUM formulas
        sum_count = 0
        bold_count = 0
        border_count = 0
        for c in range(2, 6):  # B=2, C=3, D=4, E=5
            cell = ws.cell(row=10, column=c)
            val = cell.value
            col_letter = chr(64 + c)  # B, C, D, E

            # Check formula
            if val is not None and isinstance(val, str):
                normalized = val.upper().replace(" ", "")
                expected = f"=SUM({col_letter}4:{col_letter}8)"
                if normalized == expected:
                    sum_count += 1

            # Check bold
            if cell.font.bold:
                bold_count += 1

            # Check top border
            if cell.border.top.style is not None:
                border_count += 1

        if sum_count == 4:
            comp5_score += 0.08
            print(f"  PASS: B10:E10 all have SUM formulas")
        else:
            print(f"  FAIL: Only {sum_count}/4 SUM formulas in B10:E10")

        if bold_count == 4:
            comp5_score += 0.04
            print(f"  PASS: B10:E10 all bold")
        else:
            print(f"  FAIL: Only {bold_count}/4 bold cells in B10:E10")

        if border_count >= 3:
            comp5_score += 0.03
            print(f"  PASS: B10:E10 have top borders")
        else:
            print(f"  FAIL: Only {border_count}/4 cells with top border in B10:E10")

        total_score += comp5_score
        print(f"  Component 5 total: {comp5_score:.2f}/0.20 pts")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Thick borders around matrix (0.10 points)
    # Check thick outer borders on the matrix area (rows 3-8, columns A-H)
    try:
        thick_border_count = 0
        checks_total = 0

        # Top edge: row 3 should have thick top border
        for c in range(1, 9):
            checks_total += 1
            if ws.cell(row=3, column=c).border.top.style in ('thick', 'medium'):
                thick_border_count += 1

        # Bottom edge: row 8 should have thick bottom border
        for c in range(1, 9):
            checks_total += 1
            if ws.cell(row=8, column=c).border.bottom.style in ('thick', 'medium'):
                thick_border_count += 1

        # Left edge: column A rows 3-8 should have thick left border
        for r in range(3, 9):
            checks_total += 1
            if ws.cell(row=r, column=1).border.left.style in ('thick', 'medium'):
                thick_border_count += 1

        # Right edge: column H rows 3-8 should have thick right border
        for r in range(3, 9):
            checks_total += 1
            if ws.cell(row=r, column=8).border.right.style in ('thick', 'medium'):
                thick_border_count += 1

        ratio = thick_border_count / checks_total if checks_total > 0 else 0
        if ratio >= 0.7:
            print(f"PASS: Component 6 — Thick borders around matrix ({thick_border_count}/{checks_total} edges) (0.10 pts)")
            total_score += 0.10
        elif ratio >= 0.4:
            partial = round(0.10 * ratio, 2)
            print(f"PARTIAL: Component 6 — Some thick borders ({thick_border_count}/{checks_total}) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — Insufficient thick borders ({thick_border_count}/{checks_total})")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point — persist app state then verify
def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
