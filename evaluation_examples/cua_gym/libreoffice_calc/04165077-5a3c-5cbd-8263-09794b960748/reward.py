"""
Reward Script: Build a formatted employee training matrix with skill level indicators and certification tracking.
Task ID: calc_gpm_094
Domain: libreoffice_calc
Scoring:
  Component 1: Merged title row A1:J1 with formatting (0.15)
  Component 2: Skill cells converted from numbers to B/I/A/E letters (0.20)
  Component 3: Skill cells have color-coded fills (0.15)
  Component 4: Row 15 Skill Gap Analysis with COUNTIF formulas in % format (0.15)
  Component 5: Row 17 Legend row present (0.10)
  Component 6: Chart present (stacked bar) (0.15)
  Component 7: Borders applied to data cells (0.10)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_094'

# Mapping: original number -> expected letter
LEVEL_MAP = {1: 'B', 2: 'I', 3: 'A', 4: 'E'}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Training' not in wb.sheetnames:
        print("FAIL: 'Training' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Training']

    # Component 1: Merged title A1:J1 with bold 14pt centered white on purple (0.15 pts)
    try:
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'J1' in r for r in merged_ranges)
        a1 = ws['A1']
        title_ok = (a1.value is not None and 'Training' in str(a1.value) and 'Skills' in str(a1.value))
        bold_ok = (a1.font.bold is True)
        size_ok = (a1.font.size is not None and a1.font.size >= 13)
        align_ok = (a1.alignment.horizontal == 'center')
        # Purple fill check (should be around 4B0082)
        try:
            fill_rgb = a1.fill.fgColor.rgb
            fill_ok = (a1.fill.fill_type == 'solid' and fill_rgb is not None)
        except Exception:
            fill_ok = False

        if has_merge and title_ok and bold_ok and size_ok and align_ok and fill_ok:
            print(f"PASS: Component 1 — Merged title with formatting (0.15 pts)")
            total_score += 0.15
        else:
            details = f"merge={has_merge}, title={title_ok}, bold={bold_ok}, size={size_ok}, align={align_ok}, fill={fill_ok}"
            print(f"FAIL: Component 1 — Title formatting incomplete: {details}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Skill cells B4:H13 contain letter abbreviations B/I/A/E instead of numbers (0.20 pts)
    try:
        valid_letters = {'B', 'I', 'A', 'E'}
        letter_count = 0
        total_cells = 0
        for row in range(4, 14):
            for col in range(2, 9):  # B=2 to H=8
                total_cells += 1
                val = ws.cell(row=row, column=col).value
                if isinstance(val, str) and val.strip() in valid_letters:
                    letter_count += 1

        # Need majority of cells to have letters (at least 80%)
        ratio = letter_count / total_cells if total_cells > 0 else 0
        if ratio >= 0.8:
            print(f"PASS: Component 2 — Skill cells have letter abbreviations ({letter_count}/{total_cells}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Only {letter_count}/{total_cells} skill cells have B/I/A/E letters (ratio={ratio:.2f})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Skill cells have color-coded fills (0.15 pts)
    # Check that at least some skill cells have non-default fills
    try:
        colored_count = 0
        total_skill_cells = 0
        for row in range(4, 14):
            for col in range(2, 9):
                total_skill_cells += 1
                cell = ws.cell(row=row, column=col)
                try:
                    fill_rgb = cell.fill.fgColor.rgb
                    if fill_rgb is not None and fill_rgb != '00000000' and cell.fill.fill_type == 'solid':
                        colored_count += 1
                except Exception:
                    pass

        ratio = colored_count / total_skill_cells if total_skill_cells > 0 else 0
        if ratio >= 0.7:
            print(f"PASS: Component 3 — Skill cells have color fills ({colored_count}/{total_skill_cells}) (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — Only {colored_count}/{total_skill_cells} cells have color fills (ratio={ratio:.2f})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row 15 has Skill Gap Analysis with COUNTIF formulas and % format (0.15 pts)
    try:
        a15_val = ws.cell(row=15, column=1).value
        has_label = (a15_val is not None and 'Skill Gap' in str(a15_val))

        formula_count = 0
        pct_format_count = 0
        for col in range(2, 9):  # B15:H15
            cell = ws.cell(row=15, column=col)
            val = cell.value
            if val is not None and isinstance(val, str) and 'COUNTIF' in val.upper():
                formula_count += 1
            if cell.number_format is not None and '%' in str(cell.number_format):
                pct_format_count += 1

        if has_label and formula_count >= 5 and pct_format_count >= 5:
            print(f"PASS: Component 4 — Skill Gap Analysis row with {formula_count} COUNTIF formulas, {pct_format_count} % formatted (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — label={has_label}, formulas={formula_count}/7, pct_fmt={pct_format_count}/7")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Row 17 Legend present with B/I/A/E entries (0.10 pts)
    try:
        a17_val = ws.cell(row=17, column=1).value
        has_legend_label = (a17_val is not None and 'Legend' in str(a17_val))

        # Check that legend row contains B, I, A, E labels
        legend_letters = set()
        for col in range(2, 11):
            val = ws.cell(row=17, column=col).value
            if isinstance(val, str) and val.strip() in {'B', 'I', 'A', 'E'}:
                legend_letters.add(val.strip())

        if has_legend_label and len(legend_letters) >= 3:
            print(f"PASS: Component 5 — Legend row with {len(legend_letters)} level indicators (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — legend_label={has_legend_label}, letters_found={legend_letters}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Chart present (stacked bar type preferred) (0.15 pts)
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            # Check it's a bar chart with stacked grouping
            chart_type = getattr(chart, 'type', None)
            chart_grouping = getattr(chart, 'grouping', None)
            is_bar_type = (chart_type in ('bar', 'col'))
            is_stacked = (chart_grouping == 'stacked')

            if is_bar_type and is_stacked:
                print(f"PASS: Component 6 — Stacked bar chart found (type={chart_type}, grouping={chart_grouping}) (0.15 pts)")
                total_score += 0.15
            elif len(charts) >= 1:
                # Partial: chart exists but not exact type
                print(f"PARTIAL: Component 6 — Chart exists but type={chart_type}, grouping={chart_grouping} (0.08 pts)")
                total_score += 0.08
            else:
                print(f"FAIL: Component 6 — No chart found")
        else:
            print(f"FAIL: Component 6 — No charts in worksheet")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Borders applied to data area (0.10 pts)
    try:
        border_count = 0
        check_cells = 0
        # Check a sample of cells in the data area
        for row in [3, 4, 7, 10, 13]:
            for col in [1, 2, 5, 8, 10]:
                check_cells += 1
                cell = ws.cell(row=row, column=col)
                if cell.border.left.style is not None or cell.border.top.style is not None:
                    border_count += 1

        ratio = border_count / check_cells if check_cells > 0 else 0
        if ratio >= 0.6:
            print(f"PASS: Component 7 — Borders applied ({border_count}/{check_cells} sampled cells) (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 7 — Only {border_count}/{check_cells} sampled cells have borders (ratio={ratio:.2f})")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
