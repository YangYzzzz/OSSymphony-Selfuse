"""
Initial Setup: Employee Training Matrix - raw data without formatting
Task ID: calc_gpm_094
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_094'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Training'

    # Row 1: Title text only (no merge, no formatting - that's the task)
    ws['A1'] = 'Employee Training & Skills Matrix'

    # Row 3: Headers (plain text, no formatting)
    headers = ['Employee', 'Excel', 'Python', 'SQL', 'Tableau',
               'Project Mgmt', 'Communication', 'Leadership',
               'Certifications', 'Training Hours']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Rows 4-13: 10 employees with raw numeric data
    employees = [
        ['Sarah Chen',       3, 4, 3, 2, 3, 4, 3, 4, 52],
        ['Marcus Johnson',   2, 3, 4, 3, 2, 3, 2, 2, 38],
        ['Elena Rodriguez',  4, 2, 3, 4, 3, 3, 3, 5, 65],
        ['James Kim',        3, 3, 2, 1, 4, 2, 4, 3, 44],
        ['Priya Patel',      1, 4, 4, 3, 2, 3, 2, 1, 28],
        ['David Thompson',   4, 1, 2, 2, 3, 4, 3, 3, 71],
        ['Aisha Williams',   2, 3, 3, 4, 1, 2, 1, 0, 15],
        ['Carlos Mendez',    3, 2, 1, 3, 4, 3, 4, 4, 58],
        ['Yuki Tanaka',      4, 4, 3, 2, 3, 4, 3, 2, 46],
        ['Rachel Foster',    1, 2, 4, 1, 2, 1, 2, 1, 22],
    ]

    for r, emp in enumerate(employees, 4):
        ws.cell(row=r, column=1, value=emp[0])   # Employee name
        for c in range(1, 8):                      # Skill levels B-H (cols 2-8)
            ws.cell(row=r, column=c + 1, value=emp[c])
        ws.cell(row=r, column=9, value=emp[8])    # Certifications
        ws.cell(row=r, column=10, value=emp[9])   # Training Hours

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col_letter].width = 15

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
