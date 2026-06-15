"""
Reward Script: Statistical Analysis Workbook
Task ID: calc_gpm_028
Domain: libreoffice_calc
Scoring:
  C1: Merged header A1:K1 with title and formatting (0.15)
  C2: Descriptive Statistics section F3:G14 with formulas (0.25)
  C3: Five Number Summary section F16:G21 (0.15)
  C4: Frequency Table section A22:D27 with COUNTIFS (0.15)
  C5: Two charts present (histogram + summary chart) (0.15)
  C6: Conditional formatting data bars on frequency & box plot (0.10)
  C7: Mean cell G5 highlighted yellow (0.05)
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_028'


def persist_app_state(domain):
    """Save any unsaved LibreOffice changes."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            import time
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet 'StatAnalysis' must exist
    if 'StatAnalysis' not in wb.sheetnames:
        print("FAIL: Sheet 'StatAnalysis' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['StatAnalysis']

    # Component 1: Merged header A1:K1 with title text and formatting (0.15 points)
    try:
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        has_merge = any('A1' in mr and 'K1' in mr for mr in merged_ranges)
        a1_val = ws['A1'].value
        has_title = (a1_val is not None and 'Statistical Analysis Report' in str(a1_val))
        a1_bold = ws['A1'].font.bold
        a1_size = ws['A1'].font.size

        if has_merge and has_title and a1_bold and a1_size and a1_size >= 14:
            print(f"PASS: Component 1 — Merged header A1:K1 with title '{a1_val}', bold, size {a1_size} (0.15 pts)")
            total_score += 0.15
        else:
            details = f"merge={has_merge}, title_present={has_title}, bold={a1_bold}, size={a1_size}"
            print(f"FAIL: Component 1 — Header incomplete: {details}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Descriptive Statistics section F3:G14 with stat formulas (0.25 points)
    try:
        # Check headers
        f3_val = ws['F3'].value
        g3_val = ws['G3'].value
        has_headers = (f3_val is not None and 'Statistic' in str(f3_val)
                       and g3_val is not None and 'Value' in str(g3_val))

        # Check that stat labels exist (N, Mean, Median, Std Dev, Min, Max, etc.)
        required_stats = ['N', 'Mean', 'Median', 'Std Dev', 'Min', 'Max', 'Range']
        found_stats = 0
        for r in range(4, 15):
            label = ws.cell(row=r, column=6).value
            formula = ws.cell(row=r, column=7).value
            if label is not None and formula is not None:
                label_str = str(label).strip()
                formula_str = str(formula).strip().upper()
                # Check if label matches a required stat AND has a formula
                for stat in required_stats:
                    if stat.upper() in label_str.upper() and formula_str.startswith('='):
                        found_stats += 1
                        break

        # Check number format (3 decimals = 0.000)
        has_decimal_fmt = False
        for r in range(4, 15):
            nf = ws.cell(row=r, column=7).number_format
            if nf and '0.000' in str(nf):
                has_decimal_fmt = True
                break

        sub_score = 0.0
        if has_headers:
            sub_score += 0.05
        if found_stats >= 5:
            sub_score += 0.10
        if found_stats >= 7:
            sub_score += 0.05
        if has_decimal_fmt:
            sub_score += 0.05

        if sub_score > 0:
            print(f"PASS: Component 2 — Descriptive Stats: headers={has_headers}, stats_found={found_stats}/7, decimal_fmt={has_decimal_fmt} ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 2 — Descriptive Stats section missing or incomplete")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Five Number Summary section F16:G21 (0.15 points)
    try:
        # Check for five number summary labels: Min, Q1, Median, Q3, Max
        fns_labels = ['Min', 'Q1', 'Median', 'Q3', 'Max']
        found_fns = 0
        for r in range(16, 23):
            label = ws.cell(row=r, column=6).value
            val = ws.cell(row=r, column=7).value
            if label is not None and val is not None:
                label_str = str(label).strip()
                for fns in fns_labels:
                    if fns.upper() in label_str.upper():
                        # Verify it has a formula or value
                        val_str = str(val).strip()
                        if val_str.startswith('=') or isinstance(val, (int, float)):
                            found_fns += 1
                            break

        if found_fns >= 5:
            print(f"PASS: Component 3 — Five Number Summary complete ({found_fns}/5 measures) (0.15 pts)")
            total_score += 0.15
        elif found_fns >= 3:
            partial = 0.08
            print(f"PARTIAL: Component 3 — Five Number Summary partial ({found_fns}/5 measures) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Five Number Summary missing or incomplete ({found_fns}/5)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Frequency Table A22:D27 with COUNTIFS (0.15 points)
    try:
        # Check for frequency table headers around row 22
        has_freq_header = False
        freq_row_start = None
        for r in range(20, 30):
            a_val = ws.cell(row=r, column=1).value
            if a_val is not None and 'Bin' in str(a_val):
                has_freq_header = True
                freq_row_start = r
                break

        # Check for COUNTIFS formulas in frequency column
        countifs_found = 0
        if freq_row_start is not None:
            for r in range(freq_row_start + 1, freq_row_start + 7):
                c_val = ws.cell(row=r, column=3).value
                if c_val is not None and 'COUNTIF' in str(c_val).upper():
                    countifs_found += 1

        # Check bin labels exist
        bins_found = 0
        if freq_row_start is not None:
            for r in range(freq_row_start + 1, freq_row_start + 7):
                a_val = ws.cell(row=r, column=1).value
                if a_val is not None:
                    bins_found += 1

        sub_score = 0.0
        if has_freq_header and bins_found >= 4:
            sub_score += 0.05
        if countifs_found >= 3:
            sub_score += 0.05
        if countifs_found >= 5:
            sub_score += 0.05

        if sub_score > 0:
            print(f"PASS: Component 4 — Frequency Table: header={has_freq_header}, bins={bins_found}, countifs={countifs_found} ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 4 — Frequency Table missing")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Charts present (histogram + summary chart) (0.15 points)
    try:
        charts = ws._charts
        num_charts = len(charts)

        # Check for chart titled 'Distribution' (histogram)
        has_distribution = False
        has_summary_chart = False
        for chart in charts:
            title_text = None
            try:
                if chart.title and chart.title.tx and chart.title.tx.rich:
                    for p in chart.title.tx.rich.p:
                        for run in p.r:
                            if run.t:
                                title_text = run.t
            except Exception:
                pass

            if title_text:
                if 'Distribution' in title_text or 'Histogram' in title_text.title():
                    has_distribution = True
                if 'Five' in title_text or 'Summary' in title_text or 'Box' in title_text:
                    has_summary_chart = True

        sub_score = 0.0
        if num_charts >= 1:
            sub_score += 0.05
        if num_charts >= 2:
            sub_score += 0.05
        if has_distribution or has_summary_chart:
            sub_score += 0.05

        if sub_score > 0:
            print(f"PASS: Component 5 — Charts: count={num_charts}, distribution={has_distribution}, summary={has_summary_chart} ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 5 — No charts found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting data bars (0.10 points)
    try:
        databar_count = 0
        for cf in ws.conditional_formatting:
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    databar_count += 1

        if databar_count >= 2:
            print(f"PASS: Component 6 — Data bars found: {databar_count} rules (0.10 pts)")
            total_score += 0.10
        elif databar_count >= 1:
            print(f"PARTIAL: Component 6 — Data bars found: {databar_count} rule (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 6 — No data bar conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # Component 7: Mean cell G5 highlighted yellow background (0.05 points)
    try:
        g5_fill = ws['G5'].fill
        fill_rgb = None
        if g5_fill and g5_fill.fgColor:
            fill_rgb = g5_fill.fgColor.rgb

        # Yellow = FFFFFF00 or similar yellow shade
        is_yellow = False
        if fill_rgb and isinstance(fill_rgb, str) and len(fill_rgb) >= 6:
            # Extract RGB portion (last 6 chars)
            rgb_part = fill_rgb[-6:] if len(fill_rgb) == 8 else fill_rgb
            # Yellow has high R, high G, low B
            try:
                r_val = int(rgb_part[0:2], 16)
                g_val = int(rgb_part[2:4], 16)
                b_val = int(rgb_part[4:6], 16)
                if r_val > 200 and g_val > 200 and b_val < 100:
                    is_yellow = True
            except ValueError:
                pass

        if is_yellow:
            print(f"PASS: Component 7 — Mean cell G5 has yellow background (fill: {fill_rgb}) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 7 — Mean cell G5 not highlighted yellow (fill: {fill_rgb})")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
