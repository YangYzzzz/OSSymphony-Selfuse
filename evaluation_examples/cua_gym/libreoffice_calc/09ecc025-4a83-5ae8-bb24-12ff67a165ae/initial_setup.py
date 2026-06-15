"""
Initial Setup: Build a comprehensive statistical analysis workbook
Task ID: calc_gpm_028
Domain: libreoffice_calc

Initial state: Raw data sheet with 15 data points. No formatting, no formulas,
no charts, no additional sections. The agent must build everything.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_028'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'StatAnalysis'

    # Section 1: Raw Data (A3:B17) - just the data, no fancy formatting
    # Headers in row 3
    ws.cell(row=3, column=1, value='ID')
    ws.cell(row=3, column=2, value='Value')

    # 15 data points
    values = [23, 45, 31, 67, 52, 38, 71, 44, 55, 29, 63, 48, 36, 59, 41]
    for i, val in enumerate(values):
        ws.cell(row=4 + i, column=1, value=i + 1)
        ws.cell(row=4 + i, column=2, value=val)

    # Set reasonable column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
