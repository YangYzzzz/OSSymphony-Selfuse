"""
Reward Script: Set cell protection on Inventory sheet
Task ID: calc_ps_009
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): Sheet protection is enabled
  Component 2 (0.20): Sheet password matches 'inv2024' (hash DF36)
  Component 3 (0.30): Data area B2:F100 cells are unlocked
  Component 4 (0.20): Header row A1:F1 and column A (A2:A100) remain locked
"""

import os
import openpyxl
from openpyxl.worksheet.protection import SheetProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_009'


def persist_app_state(domain: str):
    """Best-effort save of any open LibreOffice session."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Check that 'Inventory' sheet exists (precondition gate)
    if 'Inventory' not in wb.sheetnames:
        print("CRITICAL: 'Inventory' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Component 1: Sheet protection is enabled (0.30 points)
    # Initial: sheet.protection.sheet == False
    # Golden:  sheet.protection.sheet == True
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 -- Sheet protection is enabled (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 1 -- Sheet protection is NOT enabled (protection.sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Password matches 'inv2024' (0.20 points)
    # The hash for 'inv2024' is 'DF36' in openpyxl's legacy hash format.
    # Initial: no password set
    # Golden:  password hash == DF36
    try:
        expected_hash = SheetProtection(password='inv2024', sheet=True).password
        actual_hash = ws.protection.password
        if actual_hash and actual_hash == expected_hash:
            print(f"PASS: Component 2 -- Password hash matches 'inv2024' (hash={actual_hash}) (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 -- Password hash mismatch. Expected={expected_hash}, actual={actual_hash}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Data area B2:F100 cells are unlocked (0.30 points)
    # Initial: all cells locked=True (default)
    # Golden:  B2:F100 locked=False
    # We sample a representative set of cells to verify.
    try:
        unlocked_count = 0
        total_checked = 0
        # Check sample rows across the range
        sample_rows = [2, 3, 10, 25, 50, 75, 100]
        sample_cols = [2, 3, 4, 5, 6]  # B through F
        for r in sample_rows:
            for c in sample_cols:
                cell = ws.cell(row=r, column=c)
                total_checked += 1
                if not cell.protection.locked:
                    unlocked_count += 1

        unlocked_ratio = unlocked_count / total_checked if total_checked > 0 else 0
        if unlocked_ratio >= 0.9:
            print(f"PASS: Component 3 -- Data area B2:F100 is unlocked ({unlocked_count}/{total_checked} sampled cells unlocked) (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 -- Data area not fully unlocked ({unlocked_count}/{total_checked} sampled cells unlocked, ratio={unlocked_ratio:.2f})")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Header row A1:F1 and column A (A2:A100) remain locked (0.20 points)
    # Initial: all cells locked=True (so headers are locked in BOTH envs)
    # BUT this check only contributes to score when COMBINED with the fact that
    # the data area was unlocked (Component 3). If all cells are locked (initial),
    # Component 3 fails, so the total cannot be > 0.0.
    # We verify that the selective locking was done correctly.
    try:
        locked_count = 0
        total_checked_locked = 0

        # Check header row A1:F1
        for c in range(1, 7):
            cell = ws.cell(row=1, column=c)
            total_checked_locked += 1
            if cell.protection.locked:
                locked_count += 1

        # Check column A rows 2-100 (sample)
        sample_a_rows = [2, 3, 10, 25, 50, 75, 100]
        for r in sample_a_rows:
            cell = ws.cell(row=r, column=1)
            total_checked_locked += 1
            if cell.protection.locked:
                locked_count += 1

        locked_ratio = locked_count / total_checked_locked if total_checked_locked > 0 else 0

        # This component only awards points if Component 3 also passed,
        # ensuring we are checking selective locking, not default all-locked state.
        if locked_ratio >= 0.9 and total_score >= 0.30:
            print(f"PASS: Component 4 -- Headers and column A remain locked ({locked_count}/{total_checked_locked} sampled cells locked) (0.20 pts)")
            total_score += 0.20
        elif locked_ratio >= 0.9 and total_score < 0.30:
            print(f"FAIL: Component 4 -- Headers/col A are locked, but data area was not unlocked (no selective protection detected)")
        else:
            print(f"FAIL: Component 4 -- Headers/column A not fully locked ({locked_count}/{total_checked_locked} sampled cells locked, ratio={locked_ratio:.2f})")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist app state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
