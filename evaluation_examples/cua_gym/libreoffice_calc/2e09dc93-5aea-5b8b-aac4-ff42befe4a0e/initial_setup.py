"""
Initial Setup: Set up Inventory sheet with headers and data, all cells locked, sheet unprotected.
Task ID: calc_ps_009
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Inventory'

    # Headers in A1:F1
    headers = ['Item', 'Qty', 'Price', 'Location', 'Supplier', 'Notes']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
        ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)

    # Realistic inventory data rows 2-50
    data = [
        ['Laptop Dell XPS 15', 24, 1299.99, 'Warehouse A', 'Dell Inc.', 'Latest model, 16GB RAM'],
        ['Office Chair Ergonomic', 56, 349.50, 'Warehouse B', 'Herman Miller', 'Mesh back, adjustable arms'],
        ['USB-C Hub 7-Port', 150, 45.99, 'Warehouse A', 'Anker', 'HDMI + USB3.0 + SD'],
        ['27" Monitor 4K', 38, 529.00, 'Warehouse C', 'LG Electronics', 'IPS panel, HDR10'],
        ['Wireless Mouse MX Master', 92, 89.99, 'Warehouse A', 'Logitech', 'Bluetooth + USB receiver'],
        ['Standing Desk Frame', 18, 459.00, 'Warehouse B', 'FlexiSpot', 'Electric dual motor'],
        ['Mechanical Keyboard', 65, 129.95, 'Warehouse A', 'Keychron', 'K2 hot-swappable, brown switches'],
        ['Webcam HD 1080p', 110, 69.99, 'Warehouse C', 'Logitech', 'Auto-focus, noise-cancel mic'],
        ['Docking Station Thunderbolt', 30, 289.00, 'Warehouse A', 'CalDigit', 'TS4, 18 ports'],
        ['Noise Cancelling Headphones', 45, 299.99, 'Warehouse B', 'Sony', 'WH-1000XM5'],
        ['Portable SSD 1TB', 200, 109.99, 'Warehouse A', 'Samsung', 'T7 Shield, USB 3.2'],
        ['Ethernet Cable Cat6 10ft', 500, 8.99, 'Warehouse C', 'Cable Matters', 'Snagless, gold-plated'],
        ['Power Strip Surge Protector', 75, 34.99, 'Warehouse B', 'Belkin', '12-outlet, 4000J'],
        ['Laptop Stand Aluminum', 60, 54.99, 'Warehouse A', 'Rain Design', 'mStand, silver'],
        ['Desk Lamp LED', 40, 79.99, 'Warehouse C', 'BenQ', 'ScreenBar, auto-dimming'],
        ['Wireless Charger Pad', 120, 29.99, 'Warehouse A', 'Anker', '15W fast charge, Qi'],
        ['Cable Management Kit', 85, 19.99, 'Warehouse B', 'J Channel', 'Adhesive raceway, 6-pack'],
        ['Whiteboard 48x36', 12, 189.00, 'Warehouse C', '3M', 'Porcelain, magnetic'],
        ['Printer Laser B&W', 15, 249.99, 'Warehouse B', 'Brother', 'HL-L2370DW, duplex'],
        ['Toner Cartridge TN760', 80, 54.99, 'Warehouse B', 'Brother', 'High yield, 3000 pages'],
        ['Paper A4 Ream 500', 300, 12.99, 'Warehouse C', 'HP', 'Everyday, 92 bright'],
        ['Sticky Notes 3x3 12-Pack', 150, 11.49, 'Warehouse C', '3M Post-it', 'Assorted colors'],
        ['Ballpoint Pen 12-Pack', 200, 8.99, 'Warehouse C', 'Pilot', 'G2 retractable, blue'],
        ['File Folders Legal 100-Pack', 45, 24.99, 'Warehouse C', 'Smead', 'Manila, 1/3 cut tab'],
        ['Ergonomic Wrist Rest', 70, 16.99, 'Warehouse A', 'Gimars', 'Memory foam, keyboard'],
        ['HDMI Cable 6ft', 180, 9.99, 'Warehouse A', 'AmazonBasics', 'High speed, 4K@60Hz'],
        ['Wireless Presenter Remote', 35, 32.99, 'Warehouse B', 'Logitech', 'R500s, red laser'],
        ['Monitor Arm Single', 25, 109.99, 'Warehouse B', 'Ergotron', 'LX desk mount, tall pole'],
        ['Anti-Fatigue Mat', 30, 59.99, 'Warehouse B', 'CumulusPro', '20x32, beveled edges'],
        ['USB Flash Drive 64GB', 250, 11.99, 'Warehouse A', 'SanDisk', 'Ultra Flair, USB 3.0'],
        ['Screen Privacy Filter 27"', 20, 44.99, 'Warehouse C', '3M', 'Gold, frameless'],
        ['Desk Organizer Mesh', 55, 15.99, 'Warehouse C', 'SimpleHouseware', '6-compartment'],
        ['Compressed Air Duster 3-Pack', 100, 18.99, 'Warehouse C', 'Falcon', '10oz cans'],
        ['Laptop Bag 15.6"', 40, 39.99, 'Warehouse B', 'Targus', 'Citylite, water-resistant'],
        ['Presentation Clicker', 28, 27.99, 'Warehouse B', 'Kensington', 'Green laser, 100ft range'],
        ['Bluetooth Speaker Portable', 50, 49.99, 'Warehouse A', 'JBL', 'Clip 4, waterproof'],
        ['Document Scanner Portable', 10, 299.00, 'Warehouse C', 'Fujitsu', 'ScanSnap iX1300'],
        ['Desk Phone IP', 22, 139.99, 'Warehouse B', 'Polycom', 'VVX 250, PoE'],
        ['UPS Battery Backup 1500VA', 8, 199.99, 'Warehouse A', 'CyberPower', 'CP1500PFCLCD, sine wave'],
        ['Label Maker', 35, 39.99, 'Warehouse C', 'Brother', 'PT-D210, QWERTY keyboard'],
        ['Webcam Light Ring 10"', 60, 24.99, 'Warehouse A', 'Neewer', 'Dimmable, 3 modes'],
        ['Mouse Pad XL', 90, 14.99, 'Warehouse A', 'Corsair', 'MM350 extended, cloth'],
        ['Headset USB Wired', 55, 59.99, 'Warehouse B', 'Jabra', 'Evolve2 30, noise cancel'],
        ['Smart Power Adapter', 40, 22.99, 'Warehouse A', 'TP-Link', 'Kasa Mini, energy monitor'],
        ['Binder Clips Assorted 60-Pack', 120, 6.99, 'Warehouse C', 'Officemate', '3 sizes'],
        ['Tape Dispenser Desktop', 45, 7.49, 'Warehouse C', 'Scotch', 'Heavy duty, non-slip'],
        ['Calendar Desk Pad 2025', 30, 9.99, 'Warehouse C', 'AT-A-GLANCE', '22x17, ruled blocks'],
        ['Fire Extinguisher ABC 5lb', 6, 49.99, 'Warehouse B', 'Kidde', 'Rechargeable, wall mount'],
        ['First Aid Kit 250-Piece', 10, 29.99, 'Warehouse B', 'Johnson & Johnson', 'OSHA compliant'],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 30

    # All cells are locked by default in openpyxl - no need to explicitly set
    # Sheet is NOT protected (this is the default)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
