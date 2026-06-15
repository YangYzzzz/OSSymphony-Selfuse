"""
Initial Setup: Add custom header/footer to Monthly Report sheet
Task ID: calc_mcp_069
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_069'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Monthly Report ---
    ws = wb.active
    ws.title = 'Monthly Report'

    # Headers
    headers = ['Month', 'Revenue', 'Expenses', 'Net Profit', 'Margin %', 'Headcount']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Monthly financial data (Jan-Dec 2025)
    data = [
        ['January',   128450.00, 87230.00,  41220.00,  32.1, 45],
        ['February',  134780.00, 91450.00,  43330.00,  32.1, 46],
        ['March',     142300.00, 95680.00,  46620.00,  32.8, 47],
        ['April',     138920.00, 93210.00,  45710.00,  32.9, 47],
        ['May',       151670.00, 98340.00,  53330.00,  35.2, 48],
        ['June',      147890.00, 96780.00,  51110.00,  34.6, 49],
        ['July',      155230.00, 101450.00, 53780.00,  34.6, 50],
        ['August',    149870.00, 99320.00,  50550.00,  33.7, 50],
        ['September', 162340.00, 104560.00, 57780.00,  35.6, 51],
        ['October',   158790.00, 102870.00, 55920.00,  35.2, 52],
        ['November',  167450.00, 108230.00, 59220.00,  35.4, 53],
        ['December',  173920.00, 112340.00, 61580.00,  35.4, 54],
    ]

    data_align = Alignment(horizontal='center', vertical='center')
    currency_fmt = '$#,##0.00'
    pct_fmt = '0.0%'

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = data_align
            if c in (2, 3, 4):  # Revenue, Expenses, Net Profit
                cell.number_format = currency_fmt
            elif c == 5:  # Margin %
                cell.value = val / 100.0  # Store as decimal for percentage format
                cell.number_format = pct_fmt
            elif c == 6:  # Headcount
                cell.number_format = '0'

    # Summary row
    summary_row = 14
    ws.cell(row=summary_row, column=1, value='Total').font = Font(bold=True)
    ws.cell(row=summary_row, column=1).border = thin_border
    ws.cell(row=summary_row, column=1).alignment = data_align

    for col_idx in range(2, 5):  # Revenue, Expenses, Net Profit totals
        from openpyxl.utils import get_column_letter
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=summary_row, column=col_idx,
                       value=f'=SUM({col_letter}2:{col_letter}13)')
        cell.font = Font(bold=True)
        cell.number_format = currency_fmt
        cell.border = thin_border
        cell.alignment = data_align

    # Average margin
    cell = ws.cell(row=summary_row, column=5, value='=AVERAGE(E2:E13)')
    cell.font = Font(bold=True)
    cell.number_format = pct_fmt
    cell.border = thin_border
    cell.alignment = data_align

    # Total headcount (last month)
    cell = ws.cell(row=summary_row, column=6, value='=F13')
    cell.font = Font(bold=True)
    cell.number_format = '0'
    cell.border = thin_border
    cell.alignment = data_align

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Freeze header row
    ws.freeze_panes = 'A2'

    # NO headers or footers - that's the task for the agent

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
