"""
Reward Script: Verify custom header and footer on 'Monthly Report' sheet
Task ID: calc_mcp_069
Domain: libreoffice_calc
Scoring:
  Component 1: Header left = 'Confidential' (0.25)
  Component 2: Header center = 'Monthly Financial Report' (0.25)
  Component 3: Header right = date placeholder '&D' (0.25)
  Component 4: Footer center = page number placeholder '&P' (0.25)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_069'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Monthly Report' sheet must exist
    if 'Monthly Report' not in wb.sheetnames:
        print(f"CRITICAL: Sheet 'Monthly Report' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Monthly Report']
    hf = ws.HeaderFooter
    oh = hf.oddHeader
    of = hf.oddFooter

    # Component 1: Header left section = 'Confidential' (0.25 points)
    try:
        left_text = oh.left.text if (oh and oh.left) else None
        if left_text and left_text.strip().lower() == 'confidential':
            print(f"PASS: Component 1 -- Header left = '{left_text}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 -- Expected header left 'Confidential', found: {repr(left_text)}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: Header center section = 'Monthly Financial Report' (0.25 points)
    try:
        center_text = oh.center.text if (oh and oh.center) else None
        if center_text and center_text.strip().lower() == 'monthly financial report':
            print(f"PASS: Component 2 -- Header center = '{center_text}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 2 -- Expected header center 'Monthly Financial Report', found: {repr(center_text)}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Header right section = date placeholder '&D' (0.25 points)
    try:
        right_text = oh.right.text if (oh and oh.right) else None
        if right_text and '&D' in right_text.upper().replace(' ', ''):
            print(f"PASS: Component 3 -- Header right contains date placeholder = '{right_text}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 3 -- Expected header right with '&D' date placeholder, found: {repr(right_text)}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Footer center section = page number placeholder '&P' (0.25 points)
    try:
        footer_center = of.center.text if (of and of.center) else None
        if footer_center and '&P' in footer_center.upper().replace(' ', ''):
            print(f"PASS: Component 4 -- Footer center contains page number placeholder = '{footer_center}' (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 4 -- Expected footer center with '&P' page number, found: {repr(footer_center)}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved state before verification
persist_app_state("libreoffice_calc")

# Run verification
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
