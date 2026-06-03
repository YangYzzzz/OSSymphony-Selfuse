"""
Initial Setup: Employee grade book with Scores sheet containing student data.
Task ID: calc_wf_003
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Scores ---
    ws_scores = wb.active
    ws_scores.title = 'Scores'

    # Headers
    headers = ['Name', 'HW1', 'HW2', 'Midterm', 'Project', 'Final']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws_scores.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 8 students with realistic names and scores 50-100
    students = [
        ['Sarah Chen',       78, 85, 72, 88, 91],
        ['Marcus Johnson',   92, 88, 95, 90, 87],
        ['Priya Patel',      65, 70, 58, 62, 55],
        ['James O\'Brien',   84, 79, 81, 77, 83],
        ['Aisha Mohammed',   95, 97, 93, 98, 96],
        ['Carlos Rivera',    71, 68, 75, 73, 69],
        ['Emma Watson',      88, 82, 86, 91, 89],
        ['David Kim',        53, 61, 57, 50, 64],
    ]

    data_font = Font(name='Calibri', size=11)
    for r, row_data in enumerate(students, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws_scores.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.border = thin_border
            if c >= 2:
                cell.alignment = Alignment(horizontal='center')

    # Set column widths
    ws_scores.column_dimensions['A'].width = 20
    for col_letter in ['B', 'C', 'D', 'E', 'F']:
        ws_scores.column_dimensions[col_letter].width = 12

    # Freeze header row
    ws_scores.freeze_panes = 'A2'

    # --- Sheet 2: Summary (empty, just structure) ---
    ws_summary = wb.create_sheet('Summary')

    # Only add headers - NO formulas, NO grades, NO chart
    summary_headers = ['Name', 'Weighted Avg', 'Letter Grade', 'Rank']
    for col, h in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 10

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
