"""
Reward Script: Employee Grade Book with Weighted Averages, Letter Grades, Ranks, and Pie Chart
Task ID: calc_wf_003
Domain: libreoffice_calc
Scoring:
  Component 1 - Weighted average formulas (SUMPRODUCT) in B2:B9   (0.30 pts)
  Component 2 - Letter grade formulas (nested IF) in C2:C9         (0.25 pts)
  Component 3 - RANK formulas in D2:D9                             (0.20 pts)
  Component 4 - Pie chart for grade distribution on Summary sheet  (0.25 pts)
"""

import os
import re

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_003'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Summary sheet must exist
    if 'Summary' not in wb.sheetnames:
        print("FAIL: 'Summary' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']

    # -------------------------------------------------------------------------
    # Component 1: Weighted average formulas (SUMPRODUCT) in B2:B9 (0.30 pts)
    # Task requires =SUMPRODUCT with weights {0.1, 0.15, 0.2, 0.25, 0.3}
    # Initial env: B2:B9 are empty. Golden env: B2:B9 have SUMPRODUCT formulas.
    # -------------------------------------------------------------------------
    try:
        sumproduct_count = 0
        for row in range(2, 10):  # rows 2-9 (8 students)
            val = ws.cell(row=row, column=2).value  # column B
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(" ", "")
                if "SUMPRODUCT" in val_upper:
                    sumproduct_count += 1
        if sumproduct_count == 8:
            print(f"PASS: Component 1 - All 8 SUMPRODUCT formulas found in B2:B9 (0.30 pts)")
            total_score += 0.30
        elif sumproduct_count > 0:
            partial = round(0.30 * (sumproduct_count / 8), 2)
            print(f"PARTIAL: Component 1 - {sumproduct_count}/8 SUMPRODUCT formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - No SUMPRODUCT formulas found in B2:B9")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # -------------------------------------------------------------------------
    # Component 2: Letter grade formulas (nested IF) in C2:C9 (0.25 pts)
    # Task requires IF-based letter grade assignment: A>=90, B>=80, C>=70, D>=60, F<60
    # Initial env: C2:C9 are empty. Golden env: C2:C9 have nested IF formulas.
    # -------------------------------------------------------------------------
    try:
        grade_formula_count = 0
        for row in range(2, 10):
            val = ws.cell(row=row, column=3).value  # column C
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(" ", "")
                # Accept nested IF or VLOOKUP-based grade formulas
                if "IF(" in val_upper or "VLOOKUP(" in val_upper:
                    grade_formula_count += 1
        if grade_formula_count == 8:
            print(f"PASS: Component 2 - All 8 letter grade formulas found in C2:C9 (0.25 pts)")
            total_score += 0.25
        elif grade_formula_count > 0:
            partial = round(0.25 * (grade_formula_count / 8), 2)
            print(f"PARTIAL: Component 2 - {grade_formula_count}/8 grade formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - No grade formulas found in C2:C9")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # -------------------------------------------------------------------------
    # Component 3: RANK formulas in D2:D9 (0.20 pts)
    # Task requires RANK function for student ranking.
    # Initial env: D2:D9 are empty. Golden env: D2:D9 have RANK formulas.
    # -------------------------------------------------------------------------
    try:
        rank_count = 0
        for row in range(2, 10):
            val = ws.cell(row=row, column=4).value  # column D
            if val is not None and isinstance(val, str):
                val_upper = val.upper().replace(" ", "")
                if "RANK(" in val_upper:
                    rank_count += 1
        if rank_count == 8:
            print(f"PASS: Component 3 - All 8 RANK formulas found in D2:D9 (0.20 pts)")
            total_score += 0.20
        elif rank_count > 0:
            partial = round(0.20 * (rank_count / 8), 2)
            print(f"PARTIAL: Component 3 - {rank_count}/8 RANK formulas found ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - No RANK formulas found in D2:D9")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # -------------------------------------------------------------------------
    # Component 4: Pie chart on Summary sheet (0.25 pts)
    # Task requires a pie chart showing grade distribution.
    # Initial env: no charts. Golden env: 1 pie chart.
    # -------------------------------------------------------------------------
    try:
        from openpyxl.chart import PieChart
        charts = ws._charts
        pie_charts = [ch for ch in charts if isinstance(ch, PieChart)]
        if len(pie_charts) > 0:
            print(f"PASS: Component 4 - Pie chart found on Summary sheet (0.25 pts)")
            total_score += 0.25
        else:
            # Check if any chart exists (partial credit for wrong chart type)
            if len(charts) > 0:
                print(f"PARTIAL: Component 4 - Chart found but not a PieChart (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 - No charts found on Summary sheet")
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
