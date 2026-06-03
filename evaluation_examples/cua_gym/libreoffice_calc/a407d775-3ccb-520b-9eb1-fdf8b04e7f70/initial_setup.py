"""
Initial Setup: Create API response time statistics spreadsheet
Task ID: calc_gcp_052
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_gcp_052'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'APIMetrics'

    # --- Headers ---
    headers = ['Endpoint', 'Min', 'Q1', 'Median', 'Q3', 'Max',
               'Box Bottom', 'Box Height', 'Lower Whisker', 'Upper Whisker', 'Whisker Base']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000'),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- API Endpoint Data ---
    # Columns: Endpoint, Min, Q1, Median, Q3, Max
    # Then helper columns:
    #   G: Box Bottom = Q1 (used as base for stacked bar)
    #   H: Box Height = Q3 - Q1 (IQR)
    #   I: Lower Whisker = Q1 - Min
    #   J: Upper Whisker = Max - Q3
    #   K: Whisker Base = Min (invisible base for lower whisker)
    data = [
        ['/users',    45,  120,  185,  310,  580],
        ['/orders',   60,  150,  245,  420,  750],
        ['/products', 35,   95,  160,  280,  520],
        ['/search',   50,  280,  680, 1400, 2500],
        ['/auth',     30,   55,   85,  130,  180],
    ]

    data_font = Font(name='Calibri', size=11)
    endpoint_font = Font(name='Calibri', size=11, bold=True)
    num_format = '#,##0'

    for r, row_data in enumerate(data, 2):
        endpoint, mn, q1, median, q3, mx = row_data

        ws.cell(row=r, column=1, value=endpoint).font = endpoint_font
        ws.cell(row=r, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=r, column=1).border = thin_border

        ws.cell(row=r, column=2, value=mn).font = data_font
        ws.cell(row=r, column=2).number_format = num_format
        ws.cell(row=r, column=2).alignment = header_align
        ws.cell(row=r, column=2).border = thin_border

        ws.cell(row=r, column=3, value=q1).font = data_font
        ws.cell(row=r, column=3).number_format = num_format
        ws.cell(row=r, column=3).alignment = header_align
        ws.cell(row=r, column=3).border = thin_border

        ws.cell(row=r, column=4, value=median).font = data_font
        ws.cell(row=r, column=4).number_format = num_format
        ws.cell(row=r, column=4).alignment = header_align
        ws.cell(row=r, column=4).border = thin_border

        ws.cell(row=r, column=5, value=q3).font = data_font
        ws.cell(row=r, column=5).number_format = num_format
        ws.cell(row=r, column=5).alignment = header_align
        ws.cell(row=r, column=5).border = thin_border

        ws.cell(row=r, column=6, value=mx).font = data_font
        ws.cell(row=r, column=6).number_format = num_format
        ws.cell(row=r, column=6).alignment = header_align
        ws.cell(row=r, column=6).border = thin_border

        # Helper columns
        box_bottom = q1
        box_height = q3 - q1
        lower_whisker = q1 - mn
        upper_whisker = mx - q3
        whisker_base = mn

        ws.cell(row=r, column=7, value=box_bottom).font = data_font
        ws.cell(row=r, column=7).number_format = num_format
        ws.cell(row=r, column=7).alignment = header_align

        ws.cell(row=r, column=8, value=box_height).font = data_font
        ws.cell(row=r, column=8).number_format = num_format
        ws.cell(row=r, column=8).alignment = header_align

        ws.cell(row=r, column=9, value=lower_whisker).font = data_font
        ws.cell(row=r, column=9).number_format = num_format
        ws.cell(row=r, column=9).alignment = header_align

        ws.cell(row=r, column=10, value=upper_whisker).font = data_font
        ws.cell(row=r, column=10).number_format = num_format
        ws.cell(row=r, column=10).alignment = header_align

        ws.cell(row=r, column=11, value=whisker_base).font = data_font
        ws.cell(row=r, column=11).number_format = num_format
        ws.cell(row=r, column=11).alignment = header_align

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 16
    ws.column_dimensions['K'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
