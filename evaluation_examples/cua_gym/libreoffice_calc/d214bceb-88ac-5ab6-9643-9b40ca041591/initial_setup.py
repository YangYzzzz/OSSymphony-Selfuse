"""
Initial Setup: Create a project Gantt chart spreadsheet with task data and week columns.
Task ID: calc_gg2_026
Domain: libreoffice_calc

The spreadsheet has a 'Gantt' sheet with:
- Row 1: Headers
- Row 2: Week numbers (1-12) in columns C through N
- Rows 3-20: Project tasks with Start Week (col A) and End Week (col B)
- C3:N20: Empty Gantt grid (agent must apply conditional formatting)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gg2_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Gantt"

    # --- Row 1: Headers ---
    header_font = Font(name="Arial", size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    ws.cell(row=1, column=1, value="Start Week").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align

    ws.cell(row=1, column=2, value="End Week").font = header_font
    ws.cell(row=1, column=2).fill = header_fill
    ws.cell(row=1, column=2).alignment = header_align

    for col_idx in range(3, 15):  # C=3 to N=14
        week_num = col_idx - 2  # C=1, D=2, ..., N=12
        cell = ws.cell(row=1, column=col_idx, value=f"Week {week_num}")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # --- Row 2: Week numbers in C2:N2 ---
    for col_idx in range(3, 15):
        week_num = col_idx - 2
        cell = ws.cell(row=2, column=col_idx, value=week_num)
        cell.alignment = Alignment(horizontal="center")
        cell.font = Font(name="Arial", size=10, bold=True)

    ws.cell(row=2, column=1, value="").alignment = Alignment(horizontal="center")
    ws.cell(row=2, column=2, value="").alignment = Alignment(horizontal="center")

    # --- Rows 3-20: Project task data ---
    # Realistic project tasks with start and end weeks
    tasks = [
        # (Start Week, End Week) - representing a software development project
        (1, 3),    # Requirements gathering
        (2, 4),    # Stakeholder interviews
        (3, 5),    # System architecture design
        (4, 6),    # Database schema design
        (5, 8),    # Backend API development
        (5, 7),    # Frontend wireframes
        (6, 9),    # Frontend development
        (7, 9),    # Integration testing
        (8, 10),   # User acceptance testing
        (9, 10),   # Performance optimization
        (10, 11),  # Security audit
        (10, 12),  # Documentation
        (11, 12),  # Deployment preparation
        (1, 2),    # Project kickoff
        (3, 6),    # Vendor evaluation
        (6, 8),    # Infrastructure setup
        (9, 11),   # Training materials
        (11, 12),  # Go-live and support
    ]

    for row_idx, (start_wk, end_wk) in enumerate(tasks, 3):
        ws.cell(row=row_idx, column=1, value=start_wk).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=end_wk).alignment = Alignment(horizontal="center")

    # --- Column widths ---
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    for col_letter in ["C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[col_letter].width = 10

    # --- Light grid borders on the Gantt area for visual clarity ---
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    for r in range(3, 21):
        for c in range(3, 15):
            ws.cell(row=r, column=c).border = thin_border

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
