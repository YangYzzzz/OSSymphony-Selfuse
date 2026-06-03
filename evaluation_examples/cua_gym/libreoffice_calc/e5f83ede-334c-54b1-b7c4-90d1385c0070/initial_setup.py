"""
Initial Setup: 12-month cash flow forecast spreadsheet
Task ID: calc_fin_cashflow_forecast_006
Domain: libreoffice_calc

Creates initial file with:
  - Assumptions sheet: month names + projected revenue (no formulas)
  - CashFlow sheet: headers + month names + starting balance in F1
  - All formula cells (B:E) are empty — to be filled by the agent
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_fin_cashflow_forecast_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    # ----------------------------------------------------------------
    # Sheet 1: Assumptions
    # ----------------------------------------------------------------
    ws_assumptions = wb.active
    ws_assumptions.title = 'Assumptions'

    # Month names in A1:A12
    months = [
        'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]

    # Projected revenue values in B1:B12 (realistic SaaS business growth)
    revenues = [
        120000, 135000, 142000, 148000, 155000, 163000,
        170000, 178000, 185000, 192000, 200000, 215000
    ]

    for i, (month, rev) in enumerate(zip(months, revenues), start=1):
        ws_assumptions.cell(row=i, column=1, value=month)
        ws_assumptions.cell(row=i, column=2, value=rev)

    # Set column widths for readability
    ws_assumptions.column_dimensions['A'].width = 15
    ws_assumptions.column_dimensions['B'].width = 16

    # ----------------------------------------------------------------
    # Sheet 2: CashFlow
    # ----------------------------------------------------------------
    ws_cashflow = wb.create_sheet('CashFlow')

    # Row 1 headers
    headers = ['Month', 'Revenue', 'OpEx', 'Net Cash', 'Cumulative Balance']
    for col, header in enumerate(headers, 1):
        ws_cashflow.cell(row=1, column=col, value=header)

    # A2:A13 — month names (must match Assumptions sheet for VLOOKUP)
    for i, month in enumerate(months, start=2):
        ws_cashflow.cell(row=i, column=1, value=month)

    # F1 — starting cash balance = 50000
    ws_cashflow['F1'] = 50000

    # All other cells (B2:E13) are intentionally left empty
    # The agent is expected to fill these with formulas

    # Column widths for readability
    ws_cashflow.column_dimensions['A'].width = 15
    ws_cashflow.column_dimensions['B'].width = 14
    ws_cashflow.column_dimensions['C'].width = 14
    ws_cashflow.column_dimensions['D'].width = 14
    ws_cashflow.column_dimensions['E'].width = 20
    ws_cashflow.column_dimensions['F'].width = 14

    # ----------------------------------------------------------------
    # Save
    # ----------------------------------------------------------------
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheets: {wb.sheetnames}')
    print(f'  Assumptions: 12 rows of month + revenue data')
    print(f'  CashFlow: headers set, months in A2:A13, starting balance F1=50000')
    print(f'  Formula cells (B2:E13) are empty — ready for agent to fill')

create_initial()
