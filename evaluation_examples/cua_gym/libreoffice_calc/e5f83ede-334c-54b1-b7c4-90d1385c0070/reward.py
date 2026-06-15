"""
Reward Script: 12-month cash flow forecast with VLOOKUP, OpEx, Net Cash,
               Cumulative Balance, and line chart
Task ID: calc_fin_cashflow_forecast_006
Domain: libreoffice_calc
Scoring:
  Component 1: VLOOKUP revenue formulas in B2:B13         (0.25 pts)
  Component 2: OpEx formulas (65% of revenue) in C2:C13   (0.20 pts)
  Component 3: Net Cash formulas in D2:D13                 (0.20 pts)
  Component 4: Cumulative Balance formulas in E2:E13       (0.20 pts)
  Component 5: Line chart with correct title and data ref  (0.15 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fin_cashflow_forecast_006'


def count_vlookup_formulas(ws, col, start_row, end_row):
    """Count cells with VLOOKUP referencing Assumptions in given column."""
    count = 0
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str):
            upper = val.upper().replace(' ', '')
            if 'VLOOKUP' in upper and 'ASSUMPTIONS' in upper:
                count += 1
    return count


def count_opex_formulas(ws, col, start_row, end_row):
    """Count cells with OpEx formula =Bx*0.65."""
    count = 0
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str):
            norm = val.upper().replace(' ', '')
            b_ref = f'B{row}'
            if ('0.65' in norm) and (b_ref in val.upper()):
                count += 1
    return count


def count_netcash_formulas(ws, col, start_row, end_row):
    """Count cells with Net Cash formula =Bx-Cx."""
    count = 0
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str):
            norm = val.upper().replace(' ', '')
            b_ref = f'B{row}'
            c_ref = f'C{row}'
            if (b_ref in val.upper()) and (c_ref in val.upper()) and ('-' in norm):
                count += 1
    return count


def check_e2_formula(ws):
    """Check E2 references starting balance F1 and D2."""
    val = ws['E2'].value
    if val and isinstance(val, str):
        # Strip $ for substring matching: =$F$1+D2 -> =F1+D2
        norm = val.upper().replace(' ', '').replace('$', '')
        return ('F1' in norm) and ('D2' in norm)
    return False


def count_rolling_cumulative(ws, col, start_row, end_row):
    """Count cells with rolling cumulative formula =E(prev)+D(curr)."""
    count = 0
    for row in range(start_row, end_row + 1):
        val = ws.cell(row=row, column=col).value
        if val and isinstance(val, str):
            upper = val.upper().replace(' ', '')
            prev_e = f'E{row - 1}'
            curr_d = f'D{row}'
            if (prev_e in val.upper()) and (curr_d in val.upper()) and ('+' in upper):
                count += 1
    return count


def check_line_chart_data_ref(chart):
    """Return True if chart series references E2:E13."""
    if not chart.series:
        return False
    for s in chart.series:
        try:
            ref_str = s.val.numRef.f if s.val and s.val.numRef else ''
            # Clean: remove $, quotes, spaces → e.g. CASHFLOW!E2:E13
            ref_clean = ref_str.upper().replace('$', '').replace("'", '').replace(' ', '')
            if 'E2' in ref_clean and 'E13' in ref_clean:
                return True
        except Exception:
            pass
    return False


def check_chart_title(chart):
    """Return True if chart title contains 'Cash Balance Forecast'."""
    try:
        title_text = chart.title.tx.rich.p[0].r[0].t
        return 'cash balance forecast' in title_text.lower()
    except Exception:
        pass
    # Fallback: iterate all paragraphs/runs
    try:
        for para in chart.title.tx.rich.p:
            for run in para.r:
                if run.t and 'cash balance forecast' in run.t.lower():
                    return True
    except Exception:
        pass
    return False


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify required sheets exist (precondition gate)
    if 'CashFlow' not in wb.sheetnames:
        print("CRITICAL: 'CashFlow' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['CashFlow']

    # -------------------------------------------------------------------
    # Component 1: VLOOKUP revenue formulas in B2:B13 (0.25 points)
    # Each cell should contain =VLOOKUP(Ax,Assumptions.$A$1:$B$12,2,0)
    # -------------------------------------------------------------------
    try:
        vlookup_count = count_vlookup_formulas(ws, col=2, start_row=2, end_row=13)
        if vlookup_count == 12:
            print("PASS: Component 1 — All 12 VLOOKUP formulas in B2:B13 (0.25 pts)")
            total_score += 0.25
        elif vlookup_count > 0:
            partial = round(0.25 * vlookup_count / 12, 4)
            print(f"PARTIAL: Component 1 — {vlookup_count}/12 VLOOKUP formulas in B2:B13 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No VLOOKUP formulas in B2:B13 (B2={repr(ws['B2'].value)})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -------------------------------------------------------------------
    # Component 2: OpEx formulas (65% of Revenue) in C2:C13 (0.20 points)
    # Each cell should be =Bx*0.65
    # -------------------------------------------------------------------
    try:
        opex_count = count_opex_formulas(ws, col=3, start_row=2, end_row=13)
        if opex_count == 12:
            print("PASS: Component 2 — All 12 OpEx formulas (*0.65) in C2:C13 (0.20 pts)")
            total_score += 0.20
        elif opex_count > 0:
            partial = round(0.20 * opex_count / 12, 4)
            print(f"PARTIAL: Component 2 — {opex_count}/12 OpEx formulas in C2:C13 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No OpEx formulas in C2:C13 (C2={repr(ws['C2'].value)})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # -------------------------------------------------------------------
    # Component 3: Net Cash formulas in D2:D13 (0.20 points)
    # Each cell should be =Bx-Cx
    # -------------------------------------------------------------------
    try:
        netcash_count = count_netcash_formulas(ws, col=4, start_row=2, end_row=13)
        if netcash_count == 12:
            print("PASS: Component 3 — All 12 Net Cash formulas in D2:D13 (0.20 pts)")
            total_score += 0.20
        elif netcash_count > 0:
            partial = round(0.20 * netcash_count / 12, 4)
            print(f"PARTIAL: Component 3 — {netcash_count}/12 Net Cash formulas in D2:D13 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No Net Cash formulas in D2:D13 (D2={repr(ws['D2'].value)})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # -------------------------------------------------------------------
    # Component 4: Cumulative Balance formulas in E2:E13 (0.20 points)
    # E2: =$F$1+D2  (starting balance + first month net cash)
    # E3:E13: rolling =E(prev)+D(current)
    # -------------------------------------------------------------------
    try:
        e2_ok = check_e2_formula(ws)
        rolling_count = count_rolling_cumulative(ws, col=5, start_row=3, end_row=13)

        if e2_ok and rolling_count == 11:
            print("PASS: Component 4 — E2=$F$1+D2 and E3:E13 rolling cumulative (0.20 pts)")
            total_score += 0.20
        elif e2_ok or rolling_count > 0:
            parts_ok = (1 if e2_ok else 0) + rolling_count
            partial = round(0.20 * parts_ok / 12, 4)
            print(f"PARTIAL: Component 4 — e2_ok={e2_ok}, rolling={rolling_count}/11 ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Cumulative formulas missing (E2={repr(ws['E2'].value)})")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # -------------------------------------------------------------------
    # Component 5: Line chart on CashFlow with E2:E13 data and
    #              "Cash Balance Forecast" title (0.15 points)
    # -------------------------------------------------------------------
    try:
        charts = ws._charts
        if not charts:
            print("FAIL: Component 5 — No charts found on CashFlow sheet")
        else:
            # Find the first LineChart and check its properties
            line_charts = [c for c in charts if 'Line' in type(c).__name__]
            other_charts = [c for c in charts if 'Line' not in type(c).__name__]

            if not line_charts and not other_charts:
                print("FAIL: Component 5 — No charts at all on CashFlow sheet")
            elif not line_charts:
                print(f"FAIL: Component 5 — No LineChart found; found: {[type(c).__name__ for c in other_charts]}")
            else:
                # Check first line chart
                lc = line_charts[0]
                has_correct_data = check_line_chart_data_ref(lc)
                has_correct_title = check_chart_title(lc)

                if has_correct_data and has_correct_title:
                    print("PASS: Component 5 — LineChart with E2:E13 data and 'Cash Balance Forecast' title (0.15 pts)")
                    total_score += 0.15
                elif has_correct_data and not has_correct_title:
                    print("PARTIAL: Component 5 — LineChart correct data ref, title mismatch (0.10 pts)")
                    total_score += 0.10
                elif has_correct_title and not has_correct_data:
                    print("PARTIAL: Component 5 — LineChart correct title, data ref mismatch (0.07 pts)")
                    total_score += 0.07
                elif not has_correct_data and not has_correct_title:
                    print("PARTIAL: Component 5 — LineChart found but incorrect data ref and title (0.05 pts)")
                    total_score += 0.05
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
