"""
Initial Setup: Multi-location retail weekly sales consolidation workbook
Task ID: calc_gen_multisheet_075
Domain: libreoffice_calc

Creates a workbook with 6 sheets:
  - 'Summary': empty (to be filled by the agent)
  - 'Store1' through 'Store5': each with weekly performance data
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_multisheet_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary (empty — agent must fill this) ---
    ws_summary = wb.active
    ws_summary.title = 'Summary'
    # Leave completely empty — agent must build this from scratch

    # --- Store data: realistic retail metrics ---
    # Store1: Downtown flagship
    store_data = [
        {
            'name': 'Store1',
            'location': 'Downtown',
            'total_revenue': 48750.00,
            'transactions': 312,
            'avg_transaction': 156.25,
            'returns': 2140.50,
            'net_revenue': 46609.50,
        },
        {
            'name': 'Store2',
            'location': 'Westside Mall',
            'total_revenue': 35200.00,
            'transactions': 248,
            'avg_transaction': 141.94,
            'returns': 1850.00,
            'net_revenue': 33350.00,
        },
        {
            'name': 'Store3',
            'location': 'Northgate',
            'total_revenue': 52100.00,
            'transactions': 389,
            'avg_transaction': 133.93,
            'returns': 2980.75,
            'net_revenue': 49119.25,
        },
        {
            'name': 'Store4',
            'location': 'East Side Plaza',
            'total_revenue': 29640.00,
            'transactions': 201,
            'avg_transaction': 147.46,
            'returns': 1120.00,
            'net_revenue': 28520.00,
        },
        {
            'name': 'Store5',
            'location': 'Harbor View',
            'total_revenue': 41380.00,
            'transactions': 275,
            'avg_transaction': 150.47,
            'returns': 2250.00,
            'net_revenue': 39130.00,
        },
    ]

    for store in store_data:
        ws = wb.create_sheet(store['name'])

        # Store header
        ws['A1'] = f"{store['name']} — {store['location']} Weekly Report"
        ws['A1'].font = Font(name='Calibri', size=13, bold=True)
        ws.merge_cells('A1:B1')
        ws['A1'].alignment = Alignment(horizontal='center')

        # Metric labels and values
        metrics = [
            ('Total Revenue', store['total_revenue']),
            ('Transactions', store['transactions']),
            ('Avg Transaction Value', store['avg_transaction']),
            ('Returns', store['returns']),
            ('Net Revenue', store['net_revenue']),
        ]

        for row_idx, (label, value) in enumerate(metrics, start=2):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            ws.cell(row=row_idx, column=1).font = Font(name='Calibri', size=11)
            ws.cell(row=row_idx, column=2).font = Font(name='Calibri', size=11)
            if label not in ('Transactions',):
                ws.cell(row=row_idx, column=2).number_format = '#,##0.00'

        # Column widths
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 16

    # Reorder: Summary first
    # Already first since it's the active (default) sheet

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: Summary (empty), Store1, Store2, Store3, Store4, Store5')
    print('Each store sheet: B2=Total Revenue, B3=Transactions, B4=Avg Trans, B5=Returns, B6=Net Revenue')


create_initial()
