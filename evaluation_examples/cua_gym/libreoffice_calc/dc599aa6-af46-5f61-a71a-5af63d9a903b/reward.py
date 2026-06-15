"""
Reward Script: Build master consolidation workbook with Summary sheet
Task ID: calc_gen_multisheet_075
Domain: libreoffice_calc
Scoring:
  Component 1: Summary title row (merged, bold, correct text)          0.15 pts
  Component 2: Column headers in row 2 (all 7 data headers + A2 formula) 0.15 pts
  Component 3: Store data rows 3-7 with linked formulas (5 stores)    0.25 pts
  Component 4: TOTAL row 8 with SUM/AVERAGE formulas                  0.15 pts
  Component 5: Rank (G) and vs Average (H) formula columns            0.15 pts
  Component 6: Conditional formatting on Net Revenue (F3:F7)          0.15 pts
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_gen_multisheet_075'

EXPECTED_STORES = ['Store1', 'Store2', 'Store3', 'Store4', 'Store5']
EXPECTED_HEADERS = ['Revenue', 'Transactions', 'Avg Trans', 'Returns', 'Net Revenue', 'Rank', 'vs Average']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: Summary sheet must exist and have data
    if 'Summary' not in wb.sheetnames:
        print("CRITICAL: 'Summary' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Summary']
    if ws.max_row is None or ws.max_row < 3:
        print("CRITICAL: Summary sheet is empty or has insufficient rows")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: Summary title row — A1 merged across row, bold, contains 'Weekly Performance Summary' (0.15 pts)
    try:
        a1_val = ws['A1'].value
        a1_bold = ws['A1'].font.bold

        # Check that A1 is part of a merged range
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        a1_is_in_merged = any('A1' in r for r in merged_ranges)

        # Title must contain expected text
        title_has_text = (
            a1_val is not None
            and 'Weekly Performance Summary' in str(a1_val)
        )

        if title_has_text and a1_bold and a1_is_in_merged:
            print(f"PASS: Component 1 — Title 'Weekly Performance Summary' in A1, merged={a1_is_in_merged}, bold={a1_bold} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 1 — title_text={title_has_text}, bold={a1_bold}, merged={a1_is_in_merged}, A1={repr(a1_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Column headers in row 2 — A2 formula for Best store, B2-H2 with correct header names (0.15 pts)
    try:
        a2_val = ws['A2'].value

        # A2 must have a formula that identifies the best store using INDEX/MATCH
        a2_val_str = str(a2_val) if a2_val is not None else ''
        a2_has_formula = (
            a2_val_str.startswith('=')
            and 'MATCH' in a2_val_str.upper()
            and 'INDEX' in a2_val_str.upper()
        )

        # B2-H2 must match expected header names exactly
        headers_found = 0
        for col_offset, expected in enumerate(EXPECTED_HEADERS):
            col = col_offset + 2  # B=2 through H=8
            cell_val = ws.cell(row=2, column=col).value
            if cell_val is not None and str(cell_val).strip() == expected:
                headers_found += 1

        all_headers_present = (headers_found == len(EXPECTED_HEADERS))

        if a2_has_formula and all_headers_present:
            print(f"PASS: Component 2 — A2 has INDEX/MATCH best-store formula, all {headers_found} column headers correct (0.15 pts)")
            total_score += 0.15
        elif all_headers_present:
            print(f"PARTIAL: Component 2 — All headers ok but A2 missing INDEX/MATCH formula (A2={repr(a2_val)}), awarding 0.07 pts")
            total_score += 0.07
        else:
            print(f"FAIL: Component 2 — headers_found={headers_found}/{len(EXPECTED_HEADERS)}, a2_has_formula={a2_has_formula}, A2={repr(a2_val)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Rows 3-7 contain store labels (A) and formulas linking to store sheets for B-F (0.25 pts)
    try:
        stores_with_label = 0
        stores_with_links = 0

        for row_offset, store in enumerate(EXPECTED_STORES):
            row = row_offset + 3  # rows 3-7
            store_label = ws.cell(row=row, column=1).value

            if store_label == store:
                stores_with_label += 1

            # Columns B-F must have formulas that reference the store sheet by name
            linked_cols = 0
            for col in range(2, 7):  # B=2, C=3, D=4, E=5, F=6
                cell_val = ws.cell(row=row, column=col).value
                cell_val_str = str(cell_val) if cell_val is not None else ''
                if cell_val_str.startswith('=') and store in cell_val_str:
                    linked_cols += 1

            if linked_cols >= 5:
                stores_with_links += 1

        comp3_partial = round(0.025 * stores_with_label + 0.025 * stores_with_links, 2)
        comp3_partial = min(comp3_partial, 0.20)
        if stores_with_label == 5 and stores_with_links == 5:
            print(f"PASS: Component 3 — All 5 store rows present with B-F formulas linking to store sheets (0.25 pts)")
            total_score += 0.25
        elif stores_with_label >= 1 or stores_with_links >= 1:
            print(f"PARTIAL: Component 3 — {stores_with_label}/5 labels, {stores_with_links}/5 link rows, awarding {comp3_partial} pts")
            total_score += comp3_partial
        else:
            print(f"FAIL: Component 3 — No store rows found (rows 3-7 empty or not linked)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Row 8 TOTAL row with SUM formulas for columns B,C,E,F and AVERAGE for D (0.15 pts)
    try:
        total_label = ws['A8'].value
        total_label_str = str(total_label).strip().upper() if total_label is not None else ''
        has_total_label = (total_label_str == 'TOTAL')

        # B8=SUM, C8=SUM, D8=AVERAGE, E8=SUM, F8=SUM
        agg_formulas_ok = 0
        for col_letter, func in [('B', 'SUM'), ('C', 'SUM'), ('D', 'AVERAGE'), ('E', 'SUM'), ('F', 'SUM')]:
            cell_val = ws[f'{col_letter}8'].value
            cell_val_str = str(cell_val) if cell_val is not None else ''
            if cell_val_str.startswith('=') and func in cell_val_str.upper():
                agg_formulas_ok += 1

        if has_total_label and agg_formulas_ok == 5:
            print(f"PASS: Component 4 — TOTAL row with all 5 aggregation formulas (SUM/AVERAGE) (0.15 pts)")
            total_score += 0.15
        elif has_total_label and agg_formulas_ok >= 3:
            partial_score = round(0.03 * agg_formulas_ok, 2)
            print(f"PARTIAL: Component 4 — TOTAL label ok, {agg_formulas_ok}/5 formulas, awarding {partial_score} pts")
            total_score += partial_score
        else:
            print(f"FAIL: Component 4 — total_label={repr(total_label)}, agg_formulas_ok={agg_formulas_ok}/5")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Rank column (G) and vs Average column (H) with correct formulas (0.15 pts)
    try:
        rank_formula_count = 0
        avg_formula_count = 0

        for row in range(3, 8):  # rows 3-7 (5 store rows)
            g_val = ws.cell(row=row, column=7).value  # G column = Rank
            h_val = ws.cell(row=row, column=8).value  # H column = vs Average

            g_val_str = str(g_val) if g_val is not None else ''
            if g_val_str.startswith('=') and 'RANK' in g_val_str.upper():
                rank_formula_count += 1

            h_val_str = str(h_val) if h_val is not None else ''
            if h_val_str.startswith('=') and 'AVERAGE' in h_val_str.upper():
                avg_formula_count += 1

        comp5_partial = min(round(0.015 * rank_formula_count + 0.015 * avg_formula_count, 2), 0.12)
        if rank_formula_count == 5 and avg_formula_count == 5:
            print(f"PASS: Component 5 — All 5 RANK formulas in G and all 5 vs-Average formulas in H (0.15 pts)")
            total_score += 0.15
        elif rank_formula_count >= 1 or avg_formula_count >= 1:
            print(f"PARTIAL: Component 5 — {rank_formula_count}/5 RANK, {avg_formula_count}/5 AVERAGE formulas, awarding {comp5_partial} pts")
            total_score += comp5_partial
        else:
            print(f"FAIL: Component 5 — rank_formula_count={rank_formula_count}/5, avg_formula_count={avg_formula_count}/5")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting on Net Revenue column F3:F7 with color gradient (colorScale) (0.15 pts)
    try:
        cf_rules = ws.conditional_formatting

        # Find conditional formatting that covers the F column range (rows 3-7)
        cf_covers_net_revenue = 0
        cf_has_color_scale = 0

        for cf_range in cf_rules:
            range_str = str(cf_range)
            # Check CF covers a range in column F (Net Revenue)
            if 'F3' in range_str or 'F' in range_str:
                cf_covers_net_revenue += 1
                for rule in cf_rules[cf_range]:
                    if rule.type == 'colorScale':
                        cf_has_color_scale += 1

        if cf_covers_net_revenue >= 1 and cf_has_color_scale >= 1:
            print(f"PASS: Component 6 — Conditional formatting with colorScale gradient on Net Revenue column (0.15 pts)")
            total_score += 0.15
        elif cf_covers_net_revenue >= 1:
            print(f"PARTIAL: Component 6 — CF on Net Revenue range exists but no colorScale rule, awarding 0.07 pts")
            total_score += 0.07
        else:
            all_cf_ranges = [str(r) for r in cf_rules]
            print(f"FAIL: Component 6 — No CF found on column F. All CF ranges: {all_cf_ranges}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
