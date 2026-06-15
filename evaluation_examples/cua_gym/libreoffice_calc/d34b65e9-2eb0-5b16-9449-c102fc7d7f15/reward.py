"""
FINAL REWARD SCRIPT - SUCCESS
Task: Please copy 'Raw Category Names' to 'Formatted Categories'. Strip excess whitespaces and use title case formatting (capitalize first letters, lowercase others per word). Don't modify irrelevant areas.
Generated: 2025-11-24 07:32:51
Status: success
Model: o3
Total Steps: 3
"""

import openpyxl
import os
import re


def verify_categories(file_path: str) -> float:
    """Verify that the user copied raw category names to the formatted column,
    trimmed excess whitespace, and converted them to title-case.

    Scoring breakdown (progressive – max 1.0):
        • 0.1  – workbook loads & required sheet exists
        • 0.1  – header row matches expectation
        • 0.8  – row-by-row data correctness (proportional)
    Returns a float from 0.0-1.0 and prints detailed diagnostics.
    """

    max_score = 1.0
    score = 0.0

    # ---------- 1. Load workbook ----------
    try:
        wb = openpyxl.load_workbook(file_path)
        print("✓ Workbook loaded successfully")
    except Exception as e:
        print(f"✗ Failed to load workbook: {e}")
        return 0.0  # cannot continue

    # ---------- 2. Locate sheet ----------
    sheet_name = "Categories"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print("✓ 'Categories' sheet found (0.1)")
        score += 0.1
    else:
        ws = wb.active  # fallback so we can still give partial credit
        print("✗ 'Categories' sheet not found – using active sheet")

    # ---------- 3. Verify headers ----------
    expected_headers = ["Raw Category Names", "Formatted Categories"]
    header_values = [cell.value for cell in next(ws.iter_rows(max_row=1))][:2]

    if header_values == expected_headers:
        print("✓ Header row matches expected (0.1)")
        score += 0.1
    else:
        print(f"✗ Header mismatch – found {header_values}")

    # ---------- 4. Row-by-row data verification ----------
    total_rows = 0
    correct_rows = 0
    mismatches = []

    for row_idx, (raw_val, fmt_val) in enumerate(
        ws.iter_rows(min_row=2, max_col=2, values_only=True), start=2
    ):
        # Skip completely empty trailing rows
        if raw_val is None and fmt_val is None:
            continue

        total_rows += 1

        # Expected formatted value: trim & collapse spaces, title-case result
        expected_fmt = None
        if raw_val is not None:
            cleaned = " ".join(str(raw_val).strip().split())  # trim + collapse spaces
            expected_fmt = cleaned.title()

        if expected_fmt == fmt_val:
            correct_rows += 1
        else:
            mismatches.append((row_idx, raw_val, fmt_val, expected_fmt))

    if total_rows == 0:
        print("✗ No data rows found – cannot assess")
    else:
        ratio = correct_rows / total_rows
        data_score = 0.8 * ratio  # up to 0.8 points
        score += data_score
        print(
            f"Data correctness: {correct_rows}/{total_rows} rows valid "
            f"({ratio:.0%}) – {data_score:.2f} points"
        )

        if mismatches:
            print("Mismatched rows (index, raw, formatted, expected):")
            for m in mismatches:
                print(f"  Row {m[0]}: raw='{m[1]}' | formatted='{m[2]}' | expected='{m[3]}'")

    # ---------- 5. Final score ----------
    final_score = min(max_score, round(score, 4))
    print(f"FINAL SCORE: {final_score}")
    return final_score


if __name__ == "__main__":
    # Path used by the evaluation environment – adjust if necessary.
    FILE_PATH = (
        "/home/user/please_copy_raw_category_names_to_formatted_categories_" "strip_excess_whitespaces_and_use_title_case_f.xlsx"
    )

    reward = verify_categories(FILE_PATH)
    print(f"REWARD: {reward}")

