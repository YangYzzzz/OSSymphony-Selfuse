"""
Initial Setup: Build a pivot table with Department and Team row fields
Task ID: calc_pivot_026
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_026'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OrgChart'

    # --- Headers ---
    headers = ['EmpID', 'Name', 'Department', 'Team', 'Level', 'Salary']
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Department/Team structure with target counts ---
    # Engineering=42, Product=30, Design=24, QA=24 => total=120
    dept_teams = {
        'Engineering': {
            'teams': ['Frontend', 'Backend', 'Infra'],
            'count': 42,
        },
        'Product': {
            'teams': ['Strategy', 'Analytics', 'Growth'],
            'count': 30,
        },
        'Design': {
            'teams': ['UX', 'Visual', 'Research'],
            'count': 24,
        },
        'QA': {
            'teams': ['Manual', 'Automation', 'Performance'],
            'count': 24,
        },
    }

    # Distribute employees across teams within each department
    employees = []
    emp_counter = 1

    first_names = [
        'Sarah', 'Marcus', 'Elena', 'David', 'Priya', 'James', 'Yuki', 'Carlos',
        'Amara', 'Wei', 'Olivia', 'Raj', 'Fatima', 'Noah', 'Lina', 'Tom',
        'Aisha', 'Ryan', 'Sofia', 'Kenji', 'Maya', 'Patrick', 'Zara', 'Leo',
        'Nadia', 'Tyler', 'Ines', 'Jack', 'Mia', 'Omar', 'Hannah', 'Victor',
        'Clara', 'Arjun', 'Julia', 'Ethan', 'Sana', 'Liam', 'Rosa', 'Devi',
        'Nathan', 'Kayla', 'Ahmed', 'Chloe', 'Ravi', 'Emma', 'Felix', 'Isla',
        'Samuel', 'Leila', 'Daniel', 'Ava', 'Hiroshi', 'Grace', 'Mateo', 'Lily',
        'Andre', 'Nora', 'Tariq', 'Ella', 'Ben', 'Nina', 'Isaac', 'Paula',
        'Hugo', 'Rita', 'Serge', 'Vera', 'Malik', 'Tara', 'Axel', 'Dina',
        'Jorge', 'Beth', 'Ivan', 'Lucy', 'Faisal', 'Anna', 'Soren', 'Kate',
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrov', 'Kim', 'Sharma', 'Williams', 'Tanaka', 'Rodriguez',
        'Osei', 'Zhang', 'Smith', 'Patel', 'Hassan', 'Brown', 'Fischer', 'Lee',
        'Ibrahim', 'Murphy', 'Hernandez', 'Nakamura', 'Thompson', 'O\'Brien', 'Ali', 'Martin',
        'Kowalski', 'Davis', 'Moreau', 'Taylor', 'Garcia', 'Bakker', 'Novak', 'Wilson',
        'Reeves', 'Gupta', 'Becker', 'Anderson', 'Johal', 'Clark', 'Santos', 'Rao',
        'Walker', 'Barnes', 'Youssef', 'Scott', 'Mehta', 'Hall', 'Berg', 'Young',
        'Rivera', 'Kahn', 'Cooper', 'Evans', 'Ito', 'Adams', 'Ruiz', 'Foster',
        'Duval', 'Reed', 'Sayed', 'Bell', 'Price', 'Stone', 'Wells', 'Graham',
        'Larsen', 'Costa', 'Blanc', 'Frost', 'Tudor', 'Page', 'Lund', 'Dale',
        'Vega', 'Cross', 'Popov', 'Grant', 'Malik', 'Reid', 'Holm', 'Shaw',
    ]

    levels = ['Junior', 'Mid', 'Senior', 'Staff', 'Lead', 'Principal']
    salary_ranges = {
        'Junior': (55000, 75000),
        'Mid': (70000, 95000),
        'Senior': (90000, 125000),
        'Staff': (115000, 150000),
        'Lead': (130000, 170000),
        'Principal': (155000, 200000),
    }

    name_pool = []
    for fn in first_names:
        for ln in last_names:
            name_pool.append(f'{fn} {ln}')
    random.shuffle(name_pool)

    for dept_name, info in dept_teams.items():
        teams = info['teams']
        count = info['count']
        # Distribute roughly evenly across teams, remainder to first teams
        base = count // len(teams)
        remainder = count % len(teams)
        for t_idx, team_name in enumerate(teams):
            team_count = base + (1 if t_idx < remainder else 0)
            for _ in range(team_count):
                emp_id = f'E{emp_counter:03d}'
                name = name_pool[emp_counter - 1]
                level = random.choice(levels)
                lo, hi = salary_ranges[level]
                salary = round(random.randint(lo, hi) / 500) * 500
                employees.append([emp_id, name, dept_name, team_name, level, salary])
                emp_counter += 1

    # Write employee data
    for r, row_data in enumerate(employees, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 14

    # Number format for salary
    for r in range(2, len(employees) + 2):
        ws.cell(row=r, column=6).number_format = '$#,##0'

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Total employees: {len(employees)}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
