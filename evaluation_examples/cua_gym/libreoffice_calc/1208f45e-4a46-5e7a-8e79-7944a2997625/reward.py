"""
Reward Script: Build a pivot table with Department and Team as row fields, COUNT of EmpID
Task ID: calc_pivot_026
Domain: libreoffice_calc
Scoring:
  - Component 1: Pivot sheet exists (0.15)
  - Component 2: Header row has correct columns (0.15)
  - Component 3: Department subtotals correct (0.30)
  - Component 4: Team-level detail rows present with correct counts (0.25)
  - Component 5: Grand Total row = 120 (0.15)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_026'


def persist_app_state(domain):
    """Save any unsaved LibreOffice edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A "Pivot" (or pivot-like) sheet exists beyond OrgChart (0.15 points)
    # Initial file has only OrgChart; golden adds a Pivot sheet.
    try:
        non_orgchart_sheets = [s for s in wb.sheetnames if s != 'OrgChart']
        pivot_ws = None
        if len(non_orgchart_sheets) > 0:
            # Find the sheet that contains pivot data (look for one with "Department" header)
            for sn in non_orgchart_sheets:
                ws_candidate = wb[sn]
                # Check if this sheet has pivot-like content
                if ws_candidate.max_row >= 5 and ws_candidate.max_column >= 2:
                    pivot_ws = ws_candidate
                    break
            if pivot_ws is None:
                # Fallback: use first non-OrgChart sheet
                pivot_ws = wb[non_orgchart_sheets[0]]

        if pivot_ws is not None:
            print(f"PASS: Component 1 -- Pivot sheet found: '{pivot_ws.title}' (0.15 pts)")
            total_score += 0.15
        else:
            print("FAIL: Component 1 -- No pivot sheet found (only OrgChart exists)")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # If no pivot sheet, remaining components cannot pass
    if pivot_ws is None:
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # Component 2: Header row has correct columns (0.15 points)
    # Expected: Department, Team, and a count/headcount column
    try:
        headers = []
        for c in range(1, pivot_ws.max_column + 1):
            val = pivot_ws.cell(row=1, column=c).value
            headers.append(str(val).strip().lower() if val else '')

        has_dept = any('department' in h for h in headers)
        has_team = any('team' in h for h in headers)
        has_count = any(('count' in h or 'headcount' in h or 'empid' in h) for h in headers)

        if has_dept and has_team and has_count:
            print(f"PASS: Component 2 -- Headers correct: {headers} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 -- Expected Department, Team, Count headers, found: {headers}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Department subtotals are correct (0.30 points)
    # Expected subtotals: Engineering=42, Product=30, Design=24, QA=24
    expected_subtotals = {
        'engineering': 42,
        'product': 30,
        'design': 24,
        'qa': 24
    }
    try:
        found_subtotals = {}
        # Scan all rows for subtotal rows (pattern: "<Dept> Total" or dept name with no team and a count)
        for r in range(2, pivot_ws.max_row + 1):
            col_a = pivot_ws.cell(row=r, column=1).value
            col_b = pivot_ws.cell(row=r, column=2).value
            col_c = pivot_ws.cell(row=r, column=3).value if pivot_ws.max_column >= 3 else None
            # Also handle 2-column layout where count is in col B
            count_val = col_c if pivot_ws.max_column >= 3 else col_b

            if col_a is None:
                continue

            col_a_str = str(col_a).strip().lower()

            # Check for subtotal row: "<Dept> Total" pattern with no team value
            is_subtotal = False
            dept_name = None

            if 'total' in col_a_str and 'grand' not in col_a_str:
                is_subtotal = True
                dept_name = col_a_str.replace('total', '').strip()
            elif (col_b is None or str(col_b).strip() == '') and pivot_ws.max_column >= 3:
                # Dept name with empty Team column and a count value = subtotal
                # But skip the grand total row
                if count_val is not None and 'grand' not in col_a_str and 'total' not in col_a_str:
                    # This could be a subtotal if dept name matches expected
                    for exp_dept in expected_subtotals:
                        if exp_dept in col_a_str:
                            is_subtotal = True
                            dept_name = exp_dept
                            break

            if is_subtotal and dept_name:
                for exp_dept in expected_subtotals:
                    if exp_dept in dept_name:
                        try:
                            found_subtotals[exp_dept] = int(float(count_val))
                        except (ValueError, TypeError):
                            found_subtotals[exp_dept] = None
                        break

        correct_count = 0
        for dept, expected_val in expected_subtotals.items():
            if dept in found_subtotals and found_subtotals[dept] == expected_val:
                correct_count += 1
                print(f"  PASS: {dept} subtotal = {found_subtotals[dept]}")
            elif dept in found_subtotals:
                print(f"  FAIL: {dept} subtotal = {found_subtotals[dept]}, expected {expected_val}")
            else:
                print(f"  FAIL: {dept} subtotal not found")

        # Award partial credit: 0.075 per correct subtotal (4 * 0.075 = 0.30)
        sub_score = correct_count * 0.075
        if correct_count > 0:
            print(f"PASS: Component 3 -- {correct_count}/4 department subtotals correct ({sub_score:.3f} pts)")
            total_score += sub_score
        else:
            print("FAIL: Component 3 -- No correct department subtotals found")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: Team-level detail rows present with correct counts (0.25 points)
    # Expected: 12 team rows (3 teams per 4 departments) with correct headcounts
    expected_teams = {
        ('engineering', 'frontend'): 14,
        ('engineering', 'backend'): 14,
        ('engineering', 'infra'): 14,
        ('product', 'strategy'): 10,
        ('product', 'analytics'): 10,
        ('product', 'growth'): 10,
        ('design', 'ux'): 8,
        ('design', 'visual'): 8,
        ('design', 'research'): 8,
        ('qa', 'manual'): 8,
        ('qa', 'automation'): 8,
        ('qa', 'performance'): 8,
    }
    try:
        found_teams = {}
        for r in range(2, pivot_ws.max_row + 1):
            col_a = pivot_ws.cell(row=r, column=1).value
            col_b = pivot_ws.cell(row=r, column=2).value
            col_c = pivot_ws.cell(row=r, column=3).value if pivot_ws.max_column >= 3 else None

            if col_a is None or col_b is None:
                continue

            dept_str = str(col_a).strip().lower()
            team_str = str(col_b).strip().lower()
            count_val = col_c if pivot_ws.max_column >= 3 else None

            # Skip total rows
            if 'total' in dept_str or 'grand' in dept_str:
                continue

            if count_val is not None:
                try:
                    found_teams[(dept_str, team_str)] = int(float(count_val))
                except (ValueError, TypeError):
                    pass

        team_correct = 0
        for (dept, team), expected_val in expected_teams.items():
            if (dept, team) in found_teams and found_teams[(dept, team)] == expected_val:
                team_correct += 1
            elif (dept, team) in found_teams:
                print(f"  FAIL: ({dept}, {team}) = {found_teams[(dept, team)]}, expected {expected_val}")
            # Don't print every missing team to reduce noise

        # Award partial credit proportionally: 0.25 * (correct/12)
        if team_correct > 0:
            team_score = 0.25 * (team_correct / 12)
            print(f"PASS: Component 4 -- {team_correct}/12 team rows correct ({team_score:.3f} pts)")
            total_score += team_score
        else:
            print(f"FAIL: Component 4 -- No correct team-level rows found (found teams: {list(found_teams.keys())})")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: Grand Total row = 120 (0.15 points)
    try:
        grand_total_found = False
        for r in range(2, pivot_ws.max_row + 1):
            col_a = pivot_ws.cell(row=r, column=1).value
            if col_a is not None and 'grand' in str(col_a).strip().lower() and 'total' in str(col_a).strip().lower():
                # Get the count value from last data column
                count_val = None
                for c in range(pivot_ws.max_column, 0, -1):
                    v = pivot_ws.cell(row=r, column=c).value
                    if v is not None:
                        try:
                            count_val = int(float(v))
                        except (ValueError, TypeError):
                            continue
                        break

                if count_val == 120:
                    print(f"PASS: Component 5 -- Grand Total = 120 (0.15 pts)")
                    total_score += 0.15
                    grand_total_found = True
                else:
                    print(f"FAIL: Component 5 -- Grand Total = {count_val}, expected 120")
                    grand_total_found = True
                break

        if not grand_total_found:
            print("FAIL: Component 5 -- Grand Total row not found")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
persist_app_state('libreoffice_calc')

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
