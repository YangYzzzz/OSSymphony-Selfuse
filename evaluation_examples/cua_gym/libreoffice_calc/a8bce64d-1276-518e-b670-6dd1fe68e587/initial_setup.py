"""
Initial Setup: Logistics shipping data with route names and monthly shipment counts
Task ID: osworld_calc_multi_chart_computed_010
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_010'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Shipping Data ---
    ws = wb.active
    ws.title = "Shipping Data"

    # Headers: Route, Jan, Feb, Mar, Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    headers = ['Route'] + months
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
        cell.font = Font(bold=True, name='Calibri', size=11, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # Realistic logistics route data (10 routes, monthly shipment counts)
    route_data = [
        ['Shanghai - Los Angeles',  312, 287, 335, 298, 340, 362, 389, 375, 341, 318, 297, 330],
        ['Rotterdam - New York',    245, 228, 267, 251, 273, 289, 301, 294, 278, 261, 243, 258],
        ['Singapore - Sydney',      178, 165, 192, 183, 197, 205, 218, 211, 203, 189, 172, 186],
        ['Dubai - Mumbai',          134, 121, 148, 139, 156, 163, 171, 168, 157, 144, 128, 141],
        ['Hamburg - Chicago',       198, 185, 213, 204, 221, 236, 249, 243, 227, 215, 196, 209],
        ['Tokyo - Vancouver',       156, 143, 169, 161, 175, 184, 197, 191, 178, 165, 149, 163],
        ['Hong Kong - London',      267, 249, 283, 271, 292, 308, 324, 317, 299, 284, 262, 276],
        ['Busan - Seattle',         142, 131, 157, 148, 164, 172, 186, 179, 168, 154, 137, 151],
        ['Santos - Rotterdam',       89,  81,  97,  92, 104, 111, 118, 115, 107,  98,  85,  93],
        ['Colombo - Melbourne',     113, 104, 122, 116, 128, 136, 145, 141, 132, 122, 109, 118],
    ]

    for r, row_data in enumerate(route_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.alignment = Alignment(horizontal="left")
            else:
                cell.alignment = Alignment(horizontal="center")

    # Set column widths
    ws.column_dimensions['A'].width = 28
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']:
        ws.column_dimensions[col_letter].width = 7

    # Row 1 header height
    ws.row_dimensions[1].height = 20

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
