"""
Reward Script: Calculate total shipments per route and create line chart
Task ID: osworld_calc_multi_chart_computed_010
Domain: libreoffice_calc
Scoring:
  - Component 1: Total row with label 'Total' in column A     (0.30 pts)
  - Component 2: SUM formulas in each of 12 month columns     (0.40 pts)
  - Component 3: Line chart titled 'Monthly Shipment Volume'  (0.30 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_010'


def get_chart_title_text(chart):
    """Extract plain text title from a chart title object."""
    try:
        title_obj = chart.title
        if title_obj is None:
            return None
        # Navigate the rich-text structure: title.tx.rich.p[0].r[0].t
        paragraphs = title_obj.tx.rich.p
        for para in paragraphs:
            if para.r:
                return ''.join(run.t for run in para.r if run.t)
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Add a summary (Total) row using SUM formulas for each month, and
    insert a line chart titled 'Monthly Shipment Volume'.
    """
    total_score = 0.0

    # Precondition: file must load
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # The data lives on the 'Shipping Data' sheet
    ws = None
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        break  # only one sheet expected; use the first one

    if ws is None:
        print("CRITICAL: No worksheets found in workbook")
        print("REWARD: 0.0")
        return 0.0

    # ------------------------------------------------------------------
    # Component 1: Total row label (0.30 points)
    # The last data row should have 'Total' (case-insensitive) in column A.
    # Initial file has 11 rows (header + 10 routes, no Total row).
    # Golden file has 12 rows with 'Total' in A12.
    # ------------------------------------------------------------------
    try:
        max_row = ws.max_row
        # Find a row whose column-A value is 'Total' (case-insensitive)
        total_row_idx = None
        for row_idx in range(2, max_row + 1):
            cell_val = ws.cell(row=row_idx, column=1).value
            if cell_val and str(cell_val).strip().lower() == 'total':
                total_row_idx = row_idx
                break

        if total_row_idx is not None:
            print(f"PASS: Component 1 — 'Total' label found in A{total_row_idx} (0.30 pts)")
            total_score += 0.30
        else:
            print("FAIL: Component 1 — No row with 'Total' label found in column A")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: SUM formulas covering all routes in each month column (0.40 points)
    # All 12 month columns (B-M, i.e. columns 2-13) in the Total row should
    # contain SUM formulas that span rows 2 through 11 (the 10 route rows).
    # Award 0.40 if all 12 formulas are present and correct; scale partial credit.
    # ------------------------------------------------------------------
    try:
        if total_row_idx is not None:
            correct_formulas = 0
            total_months = 12  # columns B through M

            for col_idx in range(2, 14):  # columns 2-13 = B-M
                cell_val = ws.cell(row=total_row_idx, column=col_idx).value
                if isinstance(cell_val, str):
                    # Normalize: strip whitespace and spaces around operators
                    normalized = cell_val.strip().upper().replace(' ', '')
                    # Accept =SUM(Xn:Xm) where X is the column letter
                    # We check that it is a SUM formula that references rows 2 through 11
                    import re
                    # Pattern: =SUM(<col><start>:<col><end>)
                    # We need the start row to be 2 and end row to be 11
                    m = re.match(r'^=SUM\(([A-Z]+)(\d+):([A-Z]+)(\d+)\)$', normalized)
                    if m:
                        col_start, row_start, col_end, row_end = m.groups()
                        # Column letters must match (same column), and row range 2:11
                        if col_start == col_end and int(row_start) == 2 and int(row_end) == 11:
                            correct_formulas += 1

            if correct_formulas == total_months:
                print(f"PASS: Component 2 — All 12 SUM formulas present covering rows 2:11 (0.40 pts)")
                total_score += 0.40
            elif correct_formulas > 0:
                partial = round(0.40 * correct_formulas / total_months, 3)
                print(f"PARTIAL: Component 2 — {correct_formulas}/12 correct SUM formulas ({partial} pts)")
                total_score += partial
            else:
                print("FAIL: Component 2 — No correct SUM formulas found in Total row month columns")
        else:
            print("SKIP: Component 2 — Skipped because no Total row was found (Component 1 failed)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Line chart titled 'Monthly Shipment Volume' (0.30 points)
    # The golden file has a LineChart with title 'Monthly Shipment Volume'.
    # Initial file has no charts.
    # ------------------------------------------------------------------
    try:
        all_charts = []
        for sheet in wb.worksheets:
            all_charts.extend(sheet._charts)

        if len(all_charts) == 0:
            print("FAIL: Component 3 — No charts found in workbook")
        else:
            # Check for a LineChart with the correct title
            line_charts = [c for c in all_charts if type(c).__name__ == 'LineChart']
            if len(line_charts) == 0:
                print(f"FAIL: Component 3 — Found {len(all_charts)} chart(s) but none are LineChart "
                      f"(types: {[type(c).__name__ for c in all_charts]})")
            else:
                # Check title
                target_title = 'Monthly Shipment Volume'
                matched_chart = None
                for lc in line_charts:
                    title_text = get_chart_title_text(lc)
                    if title_text and title_text.strip() == target_title:
                        matched_chart = lc
                        break

                if matched_chart is not None:
                    print(f"PASS: Component 3 — LineChart titled '{target_title}' found (0.30 pts)")
                    total_score += 0.30
                else:
                    # Partial: chart exists but title differs
                    titles_found = [get_chart_title_text(lc) for lc in line_charts]
                    print(f"FAIL: Component 3 — LineChart found but title mismatch. "
                          f"Expected '{target_title}', found: {titles_found}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point: test against the canonical artifact path on the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
