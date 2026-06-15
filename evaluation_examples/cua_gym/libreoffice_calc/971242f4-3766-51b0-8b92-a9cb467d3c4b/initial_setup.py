"""
Initial Setup: Create a spreadsheet with project time breakdown data (no charts)
Task ID: calc_chart_bar_horizontal_stacked_063
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_bar_horizontal_stacked_063'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ProjectTime ---
    ws = wb.active
    ws.title = 'ProjectTime'

    # Headers
    headers = ['Project', 'Planning', 'Development', 'Testing', 'Deployment']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Data rows (matching exact context specification)
    data = [
        ['Project Alpha', 40, 280, 80, 20],
        ['Project Beta', 60, 350, 120, 30],
        ['Project Gamma', 80, 520, 160, 40],
        ['Project Delta', 30, 180, 60, 15],
    ]
    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14

    # NO charts in initial file - agent must create the chart

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print('Sheets: ProjectTime')
    print('Data rows: 4 projects with time breakdown (no charts)')


create_initial()
