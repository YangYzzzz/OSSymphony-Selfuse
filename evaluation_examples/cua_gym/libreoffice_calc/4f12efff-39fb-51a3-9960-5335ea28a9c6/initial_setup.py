"""
Initial Setup: Column chart with uniform default color (pre-task state)
Task ID: calc_chart_column_color_each_071
Domain: libreoffice_calc

Creates a spreadsheet with regional sales data and a clustered column chart
where all bars use the same default blue color (no per-data-point formatting).
"""

import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_chart_column_color_each_071'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: ColorBars ---
    ws = wb.active
    ws.title = 'ColorBars'

    # Headers
    ws['A1'] = 'Region'
    ws['B1'] = 'Sales'
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)

    # Data rows matching context ground truth
    data = [
        ('North',   84000),
        ('South',   67000),
        ('East',    92000),
        ('West',    71000),
        ('Central', 58000),
    ]
    for row_idx, (region, sales) in enumerate(data, start=2):
        ws.cell(row=row_idx, column=1, value=region)
        ws.cell(row=row_idx, column=2, value=sales)

    # Column width for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12

    # --- Clustered column chart (all bars use default blue — no per-data-point colors) ---
    chart = BarChart()
    chart.type = 'col'       # vertical column chart
    chart.grouping = 'clustered'
    chart.title = 'Regional Sales'
    chart.y_axis.title = 'Sales ($)'
    chart.x_axis.title = 'Region'
    chart.style = 2          # default Office style (single-color series)
    chart.width = 18
    chart.height = 12

    # Data reference: B1:B6 (includes header)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=6)
    # Category labels: A2:A6
    cats = Reference(ws, min_col=1, min_row=2, max_row=6)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)

    # Place chart starting at D2
    ws.add_chart(chart, 'D2')

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
