"""
Reward Script: Hospital patient data spreadsheet - growth rate row and charts
Task ID: osworld_calc_multi_chart_computed_008
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): Growth Rate row exists in row 7 with label "Growth Rate (%)" in A7
  Component 2 (0.25): Growth rate formulas in C7:G7 compute month-over-month % change
  Component 3 (0.20): Bar chart titled "Admissions by Ward" exists
  Component 4 (0.20): Line chart titled "Admission Growth Rate (%)" exists
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_008'


def extract_chart_title(chart):
    """Extract the title text string from an openpyxl chart object."""
    try:
        title = chart.title
        if title is None:
            return None
        if isinstance(title, str):
            return title
        # Navigate the complex openpyxl title structure
        if hasattr(title, 'tx') and title.tx and hasattr(title.tx, 'rich'):
            rich = title.tx.rich
            if hasattr(rich, 'p'):
                for p in rich.p:
                    if hasattr(p, 'r'):
                        for r in p.r:
                            if hasattr(r, 't') and r.t:
                                return r.t
        # Try strRef fallback
        if hasattr(title, 'tx') and title.tx and hasattr(title.tx, 'strRef'):
            sr = title.tx.strRef
            if sr and hasattr(sr, 'v'):
                return sr.v
    except Exception:
        pass
    return None


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Locate the active/target sheet
    try:
        ws = wb.active
        print(f"INFO: Active sheet = '{ws.title}', max_row={ws.max_row}, max_col={ws.max_column}")
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print("REWARD: 0.0")
        return 0.0

    # -----------------------------------------------------------------------
    # Component 1: Growth Rate row label (0.35 points)
    # Verifies that row 7 has been added with label "Growth Rate (%)" in A7.
    # This FAILS on initial (max_row=6) and PASSES on golden (max_row=7).
    # -----------------------------------------------------------------------
    try:
        a7_value = ws.cell(row=7, column=1).value
        if a7_value is not None and str(a7_value).strip().lower() == "growth rate (%)":
            print(f"PASS: Component 1 — Growth Rate row label found in A7: '{a7_value}' (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 — Expected 'Growth Rate (%)' in A7, found: {repr(a7_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not read A7: {e}")

    # -----------------------------------------------------------------------
    # Component 2: Growth rate formulas in C7:G7 (0.25 points)
    # Verifies that columns C-G of row 7 contain growth-rate formulas
    # computing month-over-month percentage change. We check that:
    #   - At least 4 of 5 formula cells contain a formula (starts with '=')
    #   - Each formula references SUM or division by a previous month column
    # B7 is intentionally None (no month-1 base to compare from).
    # -----------------------------------------------------------------------
    try:
        formula_count = 0
        formula_with_pct_logic = 0
        for col in range(3, 8):  # columns C=3 through G=7
            cell_val = ws.cell(row=7, column=col).value
            if isinstance(cell_val, str) and cell_val.startswith('='):
                formula_count += 1
                # Check for percentage growth logic: division and *100 or %
                upper_val = cell_val.upper().replace(' ', '')
                if ('SUM' in upper_val or '/' in upper_val) and ('100' in upper_val or '%' in upper_val or '/' in upper_val):
                    formula_with_pct_logic += 1

        print(f"INFO: Component 2 — formula_count={formula_count}, formula_with_pct_logic={formula_with_pct_logic}")

        if formula_count >= 4 and formula_with_pct_logic >= 4:
            print(f"PASS: Component 2 — Growth rate formulas found in C7:G7 (0.25 pts)")
            total_score += 0.25
        elif formula_count >= 2:
            print(f"PARTIAL: Component 2 — Only {formula_count} formula cells found in C7:G7; expected 5")
        else:
            print(f"FAIL: Component 2 — Insufficient growth rate formulas in C7:G7 (found {formula_count})")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not inspect C7:G7 formulas: {e}")

    # -----------------------------------------------------------------------
    # Component 3: Bar chart titled "Admissions by Ward" (0.20 points)
    # Verifies that a bar chart with the correct title has been inserted.
    # Initial file has 0 charts; golden file has 2 charts.
    # -----------------------------------------------------------------------
    try:
        charts = ws._charts
        bar_chart_found = False
        bar_chart_title_correct = False

        for chart in charts:
            chart_type = type(chart).__name__
            if chart_type == 'BarChart':
                bar_chart_found = True
                title_text = extract_chart_title(chart)
                print(f"INFO: Component 3 — BarChart found, title='{title_text}'")
                if title_text and 'admissions by ward' in title_text.lower():
                    bar_chart_title_correct = True
                    break

        if bar_chart_found and bar_chart_title_correct:
            print(f"PASS: Component 3 — Bar chart 'Admissions by Ward' found (0.20 pts)")
            total_score += 0.20
        elif bar_chart_found:
            print(f"FAIL: Component 3 — Bar chart found but title is not 'Admissions by Ward'")
        else:
            print(f"FAIL: Component 3 — No bar chart found (total charts: {len(charts)})")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not inspect charts: {e}")

    # -----------------------------------------------------------------------
    # Component 4: Line chart titled "Admission Growth Rate (%)" (0.20 points)
    # Verifies that a line chart with the correct title has been inserted.
    # Initial file has 0 charts; golden file has 2 charts.
    # -----------------------------------------------------------------------
    try:
        charts = ws._charts
        line_chart_found = False
        line_chart_title_correct = False

        for chart in charts:
            chart_type = type(chart).__name__
            if chart_type == 'LineChart':
                line_chart_found = True
                title_text = extract_chart_title(chart)
                print(f"INFO: Component 4 — LineChart found, title='{title_text}'")
                if title_text and 'growth rate' in title_text.lower():
                    line_chart_title_correct = True
                    break

        if line_chart_found and line_chart_title_correct:
            print(f"PASS: Component 4 — Line chart 'Admission Growth Rate (%)' found (0.20 pts)")
            total_score += 0.20
        elif line_chart_found:
            print(f"FAIL: Component 4 — Line chart found but title does not contain 'Growth Rate'")
        else:
            print(f"FAIL: Component 4 — No line chart found (total charts: {len(charts)})")
    except Exception as e:
        print(f"ERROR: Component 4 — Could not inspect charts: {e}")

    # -----------------------------------------------------------------------
    # Final score
    # -----------------------------------------------------------------------
    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the VM
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
