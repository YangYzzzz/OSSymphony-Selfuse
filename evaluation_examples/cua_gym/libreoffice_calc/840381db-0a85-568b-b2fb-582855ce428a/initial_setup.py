"""
Initial Setup: Hospital patient data spreadsheet with ward admission data
Task ID: osworld_calc_multi_chart_computed_008
Domain: libreoffice_calc

Creates a spreadsheet with:
- 5 hospital wards in column A
- 6 months of admission data in columns B-G (Jan-Jun)
- NO growth rate row (to be added by agent)
- NO charts (to be added by agent)
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_multi_chart_computed_008'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hospital Admissions"

    # Header row styling
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Headers: Ward | Jan | Feb | Mar | Apr | May | Jun
    headers = ["Ward", "Jan", "Feb", "Mar", "Apr", "May", "Jun"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    # Realistic hospital ward admission data (5 wards, 6 months)
    # Ward names and monthly admissions Jan-Jun
    ward_data = [
        ["Cardiology",    142, 138, 151, 163, 155, 172],
        ["Orthopedics",    98, 104, 112, 108, 119, 127],
        ["Neurology",      87,  91,  95,  88,  97, 103],
        ["Pediatrics",    203, 198, 215, 221, 209, 234],
        ["General Surgery", 176, 182, 168, 191, 205, 198],
    ]

    data_font = Font(name="Calibri", size=11)
    ward_font = Font(name="Calibri", size=11, bold=True)

    for r, row_data in enumerate(ward_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border
            cell.alignment = center_align
            if c == 1:
                cell.font = ward_font
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.font = data_font

    # Set column widths for readability
    ws.column_dimensions["A"].width = 20
    for col_letter in ["B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[col_letter].width = 10

    # Set row heights
    ws.row_dimensions[1].height = 22
    for r in range(2, 7):
        ws.row_dimensions[r].height = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
