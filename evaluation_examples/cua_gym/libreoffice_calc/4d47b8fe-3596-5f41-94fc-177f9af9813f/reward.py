"""
Reward Script: Age calculation with DATEDIF and conditional formatting
Task ID: osworld_calc_age_calculation_datedif_006
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6 pts): Column D (D2:D15) has DATEDIF formulas calculating
                          age in years and months relative to the study date in F1.
  Component 2 (0.4 pts): Conditional formatting rule exists that highlights
                          participants over 60 years old using a formula based
                          on DATEDIF($C..,$F$1,"Y")>60 or equivalent age > 60 check.
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_age_calculation_datedif_006'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if 'Participants' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Participants' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Participants']

    # -----------------------------------------------------------------------
    # Component 1: DATEDIF formulas in D2:D15 (0.6 points)
    #
    # Task requires column D to have DATEDIF formulas calculating age as
    # "X years Y months" relative to the study date in F1.
    # A correct formula must:
    #   (a) use DATEDIF with "Y" to get full years
    #   (b) use DATEDIF with "YM" to get remaining months
    #   (c) reference the absolute cell $F$1 as the end date
    # We count how many rows (D2:D15) have a correct formula.
    # -----------------------------------------------------------------------
    try:
        data_rows = range(2, 16)   # rows 2 through 15, 14 participants
        correct_formula_count = 0
        wrong_formula_rows = []

        for row in data_rows:
            cell = ws.cell(row=row, column=4)
            val = cell.value

            if val is None:
                wrong_formula_rows.append(row)
                continue

            formula_str = str(val).upper().replace(" ", "")

            # Must contain DATEDIF(...,"Y") and DATEDIF(...,"YM")
            has_years = bool(re.search(r'DATEDIF\(', formula_str) and '"Y"' in formula_str)
            has_months = bool('"YM"' in formula_str)
            # Must reference $F$1 as study date (absolute reference to F1)
            has_f1_ref = bool('$F$1' in str(val).upper() or 'F1' in str(val).upper())

            if has_years and has_months and has_f1_ref:
                correct_formula_count += 1
            else:
                wrong_formula_rows.append(row)

        expected_count = len(data_rows)  # 14

        if correct_formula_count == expected_count:
            print(f"PASS: Component 1 — All {correct_formula_count}/{expected_count} D-column cells "
                  f"have DATEDIF 'Y'+'YM' formulas referencing F1 (0.6 pts)")
            total_score += 0.6
        elif correct_formula_count > 0:
            # Partial credit within component: award proportional fraction of 0.6
            partial = round(0.6 * correct_formula_count / expected_count, 4)
            print(f"PARTIAL: Component 1 — {correct_formula_count}/{expected_count} D-column cells "
                  f"have correct DATEDIF formulas. Missing/incorrect rows: {wrong_formula_rows}. "
                  f"Awarding {partial} pts")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — No DATEDIF formulas found in D2:D15. "
                  f"Column D appears to be empty or contains non-DATEDIF values.")

    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # -----------------------------------------------------------------------
    # Component 2: Conditional formatting rule for participants over 60 (0.4 points)
    #
    # Task requires a CF rule that highlights rows where age in years > 60.
    # The golden file uses:
    #   - Range: A2:D16 (or overlapping subset)
    #   - Type: expression (formula-based)
    #   - Formula: DATEDIF($C2,$F$1,"Y")>60 (or equivalent age > 60 check)
    # We verify:
    #   (a) At least one CF rule exists on the worksheet
    #   (b) The rule formula references an age > 60 comparison using DATEDIF or similar
    # -----------------------------------------------------------------------
    try:
        cf_rules = list(ws.conditional_formatting)
        found_age_cf = False
        cf_details = []

        for cf_range in cf_rules:
            rules_in_range = ws.conditional_formatting[cf_range]
            for rule in rules_in_range:
                formula_list = getattr(rule, 'formula', None) or []
                for f in formula_list:
                    f_upper = str(f).upper().replace(" ", "")
                    # Check for DATEDIF-based age > 60 comparison
                    has_datedif = 'DATEDIF(' in f_upper
                    has_60 = '>60' in f_upper
                    # Also accept simpler patterns like YEAR-based comparison
                    has_year_calc = ('YEAR(' in f_upper or 'INT(' in f_upper) and '>60' in f_upper
                    # Also accept formulas checking column D value
                    has_d_col_gt60 = ('$D' in f_upper or '"D"' in f_upper) and '>60' in f_upper

                    if (has_datedif and has_60) or has_year_calc or has_d_col_gt60:
                        found_age_cf = True
                        cf_details.append(f"Range: {cf_range}, Formula: {f}")
                        break
                if found_age_cf:
                    break
            if found_age_cf:
                break

        if found_age_cf:
            print(f"PASS: Component 2 — Conditional formatting rule found for age > 60. "
                  f"Details: {cf_details} (0.4 pts)")
            total_score += 0.4
        elif len(cf_rules) > 0:
            # CF rules exist but may not have correct formula
            print(f"FAIL: Component 2 — {len(cf_rules)} CF rule(s) exist but none implement "
                  f"an age > 60 check. Found formulas: "
                  f"{[getattr(r, 'formula', None) for cf in cf_rules for r in ws.conditional_formatting[cf]]}")
        else:
            print("FAIL: Component 2 — No conditional formatting rules found on the worksheet.")

    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
