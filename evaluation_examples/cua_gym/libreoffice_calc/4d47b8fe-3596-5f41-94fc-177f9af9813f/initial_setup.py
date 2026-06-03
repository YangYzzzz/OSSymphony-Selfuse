"""
Initial Setup: Age calculation study participant spreadsheet
Task ID: osworld_calc_age_calculation_datedif_006
Domain: libreoffice_calc

Creates a spreadsheet with study participant data including ID, Name, Date of Birth.
Column D has header 'Age at Study' but is empty (no formulas).
Cell F1 contains the study reference date.
No conditional formatting is present.
"""

import os
import shlex
import subprocess
import time
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'  # VM path — all scripts run on the VM
TASK_ID = 'osworld_calc_age_calculation_datedif_006'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Participants ---
    ws = wb.active
    ws.title = 'Participants'

    # Study reference date in F1
    ws['F1'] = date(2024, 6, 15)
    ws['F1'].number_format = 'YYYY-MM-DD'
    ws['E1'] = 'Study Date:'
    ws['E1'].font = Font(bold=True)

    # Headers in row 1
    headers = ['ID', 'Name', 'Date of Birth', 'Age at Study']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.alignment = Alignment(horizontal='center')

    # Realistic study participant data (Date of Birth as date objects)
    # Mix of participants including some over 60 (born before 1964-06-15)
    participants = [
        (1,  'Margaret Holloway',   date(1948, 3, 22)),   # ~76 years old
        (2,  'James Okafor',        date(1955, 11, 7)),   # ~68 years old
        (3,  'Linda Vasquez',       date(1961, 8, 30)),   # ~62 years old
        (4,  'Robert Steinberg',    date(1970, 4, 14)),   # ~54 years old
        (5,  'Patricia Nguyen',     date(1943, 12, 5)),   # ~80 years old
        (6,  'Thomas Fairbanks',    date(1967, 9, 20)),   # ~56 years old
        (7,  'Susan Delacroix',     date(1958, 2, 18)),   # ~66 years old
        (8,  'Charles Mbeki',       date(1975, 6, 3)),    # ~49 years old
        (9,  'Dorothy Yamamoto',    date(1952, 7, 29)),   # ~71 years old
        (10, 'William Petersen',    date(1980, 1, 11)),   # ~44 years old
        (11, 'Barbara Kowalski',    date(1963, 5, 16)),   # ~61 years old
        (12, 'Richard Alvarez',     date(1990, 10, 25)),  # ~33 years old
        (13, 'Sandra Ihejirika',    date(1946, 4, 8)),    # ~78 years old
        (14, 'Joseph Tremblay',     date(1985, 7, 22)),   # ~38 years old
        (15, 'Carol Blackwood',     date(1937, 9, 14)),   # ~86 years old
    ]

    for row_data in participants:
        pid, name, dob = row_data
        row_idx = pid + 1  # data starts at row 2
        ws.cell(row=row_idx, column=1, value=pid)
        ws.cell(row=row_idx, column=2, value=name)
        dob_cell = ws.cell(row=row_idx, column=3, value=dob)
        dob_cell.number_format = 'YYYY-MM-DD'
        # Column D is intentionally left EMPTY (no formula, no value)
        # This is what the agent needs to fill in

    # Column widths for readability
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Freeze top row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
