"""
Reward Script: Sort only the data section (rows 8-25, columns A-D) by Score (column C) in descending order.
Task ID: calc_dop_sort_range_065
Domain: libreoffice_calc
Scoring:
  Gate: Sheet 'ContestResults' exists and boundary rows (1-7, 26-30) are intact (fail early if not)
  Component 1 (0.5): Score column (C8:C25) sorted in descending order (highest to lowest)
  Component 2 (0.5): Row data integrity — all 4 columns (A-D) correctly paired in each row after sort
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_dop_sort_range_065'

# Ground truth: expected ordering of (Rank, Contestant, Score, Category) after sorting by Score desc
EXPECTED_DATA_ROWS = [
    (1,  'Zhang Wei',           98, 'Junior'),
    (2,  'Kwame Asante',        92, 'Senior'),
    (3,  'Mateus Ribeiro',      85, 'Open'),
    (4,  'Aisha Kamara',        81, 'Open'),
    (5,  "Bridget O'Sullivan",  79, 'Senior'),
    (6,  'Carlos Mendez',       77, 'Senior'),
    (7,  'Priya Venkataraman',  75, 'Open'),
    (8,  'Tariq Al-Rashid',     72, 'Open'),
    (9,  'Emily Thornton',      70, 'Senior'),
    (10, 'Chloe Beaumont',      68, 'Open'),
    (11, 'Nadia Petrov',        66, 'Open'),
    (12, 'Leon Okafor',         63, 'Open'),
    (13, 'Yusuf Ibrahim',       60, 'Junior'),
    (14, 'Ingrid Lassen',       58, 'Junior'),
    (15, 'Sofia Marchetti',     55, 'Senior'),
    (16, 'Hana Yoshida',        53, 'Junior'),
    (17, 'Raj Patel',           49, 'Senior'),
    (18, 'Dmitri Voronov',      42, 'Junior'),
]

# Expected boundary rows (title block rows 1-7 and summary rows 26-30)
# These are used as a GATE (fail early) — NOT as scoring components,
# because they are identical in both initial and golden files.
EXPECTED_TITLE_ROWS = {
    1: ('Regional Programming Contest 2025', None, None, None),
    2: ('Final Standings — Algorithm Track', None, None, None),
    3: ('Date: November 14, 2025   |   Venue: TechHub Convention Centre', None, None, None),
    4: ('Organized by: TechSpark Foundation   |   Chief Judge: Dr. Amara Osei', None, None, None),
    5: ('Scores are final. Ties broken by submission time.', None, None, None),
    6: (None, None, None, None),
    7: ('Rank', 'Contestant', 'Score', 'Category'),
}

EXPECTED_SUMMARY_ROWS = {
    26: (None, None, None, None),
    27: (None, 'Average Score', '=AVERAGE(C8:C25)', None),
    28: (None, 'Highest Score', '=MAX(C8:C25)', None),
    29: (None, 'Lowest Score', '=MIN(C8:C25)', None),
    30: (None, 'Total Contestants', '=COUNT(C8:C25)', None),
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Gate: Verify sheet exists
    if 'ContestResults' not in wb.sheetnames:
        print("CRITICAL: Sheet 'ContestResults' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['ContestResults']

    # Gate: Verify boundary rows are intact (title block rows 1-7 and summary rows 26-30)
    # This is a PRECONDITION gate, not a scoring component — both initial and golden have these rows.
    # A failing check here means the agent destructively modified rows it should not have touched.
    boundary_corrupted = False
    for row_num, expected_tuple in EXPECTED_TITLE_ROWS.items():
        for col_idx, expected_val in enumerate(expected_tuple, 1):
            actual_val = ws.cell(row=row_num, column=col_idx).value
            if actual_val != expected_val:
                boundary_corrupted = True
                from openpyxl.utils import get_column_letter
                print(f"GATE FAIL: Title block row {row_num} col {get_column_letter(col_idx)}: "
                      f"expected {repr(expected_val)}, found {repr(actual_val)}")

    for row_num, expected_tuple in EXPECTED_SUMMARY_ROWS.items():
        for col_idx, expected_val in enumerate(expected_tuple, 1):
            actual_val = ws.cell(row=row_num, column=col_idx).value
            if actual_val != expected_val:
                boundary_corrupted = True
                from openpyxl.utils import get_column_letter
                print(f"GATE FAIL: Summary row {row_num} col {get_column_letter(col_idx)}: "
                      f"expected {repr(expected_val)}, found {repr(actual_val)}")

    if boundary_corrupted:
        print("GATE: Boundary rows were corrupted. Returning 0.0.")
        print("REWARD: 0.0")
        return 0.0
    else:
        print("GATE: Boundary rows (1-7 and 26-30) are intact — precondition met.")

    # Read actual data rows 8-25
    actual_scores = []
    actual_rows = []
    for row in range(8, 26):
        rank = ws.cell(row=row, column=1).value
        contestant = ws.cell(row=row, column=2).value
        score = ws.cell(row=row, column=3).value
        category = ws.cell(row=row, column=4).value
        actual_scores.append(score)
        actual_rows.append((rank, contestant, score, category))

    # Component 1: Score column (C8:C25) sorted in descending order (0.5 points)
    # This FAILS on initial file (scores are in random order) and PASSES on golden (sorted desc)
    try:
        expected_scores = [row[2] for row in EXPECTED_DATA_ROWS]
        scores_in_order = True
        mismatched_scores = []
        for i, (actual_s, expected_s) in enumerate(zip(actual_scores, expected_scores)):
            if actual_s != expected_s:
                scores_in_order = False
                mismatched_scores.append(f"Row {8+i}: expected score {expected_s}, found {actual_s}")

        if scores_in_order:
            print(f"PASS: Component 1 — Score column C8:C25 sorted descending (0.5 pts)")
            print(f"  Top score in row 8: {actual_scores[0]}, bottom score in row 25: {actual_scores[-1]}")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — Score column not sorted correctly in descending order")
            for msg in mismatched_scores[:5]:
                print(f"  {msg}")
            if len(mismatched_scores) > 5:
                print(f"  ... and {len(mismatched_scores) - 5} more mismatches")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row data integrity — all 4 columns correctly paired in each row (0.5 points)
    # Verifies that columns A, B, D also moved with column C (sort applied to full row, not just score col)
    # This FAILS on initial (rows not sorted, so contestant/rank/category don't match expected positions)
    # and PASSES on golden (all columns correctly paired after sort)
    try:
        rows_correct = True
        mismatched_rows = []
        for i, (actual_row, expected_row) in enumerate(zip(actual_rows, EXPECTED_DATA_ROWS)):
            exp_rank, exp_contestant, exp_score, exp_category = expected_row
            act_rank, act_contestant, act_score, act_category = actual_row
            row_num = 8 + i
            if act_rank != exp_rank or act_contestant != exp_contestant or act_category != exp_category:
                rows_correct = False
                mismatched_rows.append(
                    f"Row {row_num}: expected (Rank={exp_rank}, Name={exp_contestant}, Score={exp_score}, Cat={exp_category}), "
                    f"found (Rank={act_rank}, Name={act_contestant}, Score={act_score}, Cat={act_category})"
                )

        if rows_correct:
            print(f"PASS: Component 2 — All 4 columns correctly paired in each row (0.5 pts)")
            print(f"  Row 8: {actual_rows[0]}")
            print(f"  Row 25: {actual_rows[-1]}")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — Row data pairing is incorrect ({len(mismatched_rows)} rows wrong)")
            for msg in mismatched_rows[:3]:
                print(f"  {msg}")
            if len(mismatched_rows) > 3:
                print(f"  ... and {len(mismatched_rows) - 3} more mismatches")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
