"""
Initial Setup: Sort data section only (rows 8-25) by Score descending
Task ID: calc_dop_sort_range_065
Domain: libreoffice_calc

Creates a spreadsheet with:
- Rows 1-5: Title block (merged cells, contest name, date)
- Row 6: empty spacer
- Row 7: headers (Rank, Contestant, Score, Category)
- Rows 8-25: 18 contestant records with scores in RANDOM order (NOT sorted)
- Row 26: empty spacer
- Rows 27-30: summary statistics
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_sort_range_065'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'ContestResults'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18

    # --- Title block: Rows 1-5 ---
    # Row 1: Contest name (merged A1:D1)
    ws.merge_cells('A1:D1')
    ws['A1'] = 'Regional Programming Contest 2025'
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 36

    # Row 2: Subtitle (merged A2:D2)
    ws.merge_cells('A2:D2')
    ws['A2'] = 'Final Standings — Algorithm Track'
    ws['A2'].font = Font(name='Calibri', size=13, italic=True, color='FFFFFF')
    ws['A2'].fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 26

    # Row 3: Date info (merged A3:D3)
    ws.merge_cells('A3:D3')
    ws['A3'] = 'Date: November 14, 2025   |   Venue: TechHub Convention Centre'
    ws['A3'].font = Font(name='Calibri', size=10, color='333333')
    ws['A3'].fill = PatternFill(start_color='FFD9E1EC', end_color='FFD9E1EC', fill_type='solid')
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 20

    # Row 4: Organizer info (merged A4:D4)
    ws.merge_cells('A4:D4')
    ws['A4'] = 'Organized by: TechSpark Foundation   |   Chief Judge: Dr. Amara Osei'
    ws['A4'].font = Font(name='Calibri', size=10, color='555555')
    ws['A4'].fill = PatternFill(start_color='FFD9E1EC', end_color='FFD9E1EC', fill_type='solid')
    ws['A4'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[4].height = 20

    # Row 5: disclaimer/note (merged A5:D5)
    ws.merge_cells('A5:D5')
    ws['A5'] = 'Scores are final. Ties broken by submission time.'
    ws['A5'].font = Font(name='Calibri', size=9, italic=True, color='777777')
    ws['A5'].fill = PatternFill(start_color='FFF5F7FA', end_color='FFF5F7FA', fill_type='solid')
    ws['A5'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 18

    # Row 6: empty spacer
    ws.row_dimensions[6].height = 8

    # --- Row 7: Data headers ---
    headers = ['Rank', 'Contestant', 'Score', 'Category']
    header_fill = PatternFill(start_color='FF3A7CA5', end_color='FF3A7CA5', fill_type='solid')
    thin = Side(style='thin', color='FFFFFF')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = header_border
    ws.row_dimensions[7].height = 22

    # --- Rows 8-25: 18 contestant records (intentionally NOT sorted by Score) ---
    # Scores range 42-98; order is random (not ascending or descending)
    contestants = [
        # (Rank, Contestant, Score, Category)
        (7,  'Priya Venkataraman',   75, 'Open'),
        (12, 'Leon Okafor',          63, 'Open'),
        (1,  'Zhang Wei',            98, 'Junior'),
        (15, 'Sofia Marchetti',      55, 'Senior'),
        (4,  'Aisha Kamara',         81, 'Open'),
        (18, 'Dmitri Voronov',       42, 'Junior'),
        (9,  'Emily Thornton',       70, 'Senior'),
        (3,  'Mateus Ribeiro',       85, 'Open'),
        (16, 'Hana Yoshida',         53, 'Junior'),
        (6,  'Carlos Mendez',        77, 'Senior'),
        (11, 'Nadia Petrov',         66, 'Open'),
        (2,  'Kwame Asante',         92, 'Senior'),
        (14, 'Ingrid Lassen',        58, 'Junior'),
        (8,  'Tariq Al-Rashid',      72, 'Open'),
        (5,  'Bridget O\'Sullivan',  79, 'Senior'),
        (13, 'Yusuf Ibrahim',        60, 'Junior'),
        (10, 'Chloe Beaumont',       68, 'Open'),
        (17, 'Raj Patel',            49, 'Senior'),
    ]

    data_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    for i, (rank, name, score, category) in enumerate(contestants):
        row = 8 + i
        row_data = [rank, name, score, category]
        # Alternating row fill
        if i % 2 == 0:
            row_fill = PatternFill(start_color='FFFAFBFC', end_color='FFFAFBFC', fill_type='solid')
        else:
            row_fill = PatternFill(start_color='FFE8EFF7', end_color='FFE8EFF7', fill_type='solid')
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = row_fill
            cell.border = data_border
            cell.font = Font(name='Calibri', size=10)
            if col == 1 or col == 3:
                cell.alignment = Alignment(horizontal='center')
            else:
                cell.alignment = Alignment(horizontal='left')
        ws.row_dimensions[row].height = 18

    # Row 26: empty spacer
    ws.row_dimensions[26].height = 8

    # --- Rows 27-30: Summary statistics ---
    summary_label_fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
    summary_data = [
        (27, 'Average Score',         '=AVERAGE(C8:C25)'),
        (28, 'Highest Score',         '=MAX(C8:C25)'),
        (29, 'Lowest Score',          '=MIN(C8:C25)'),
        (30, 'Total Contestants',     '=COUNT(C8:C25)'),
    ]
    for (row, label, formula) in summary_data:
        # Label in B
        lbl_cell = ws.cell(row=row, column=2, value=label)
        lbl_cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        lbl_cell.fill = summary_label_fill
        lbl_cell.alignment = Alignment(horizontal='right')
        # Value in C
        val_cell = ws.cell(row=row, column=3, value=formula)
        val_cell.font = Font(name='Calibri', size=10, bold=True)
        val_cell.fill = PatternFill(start_color='FFD9E1EC', end_color='FFD9E1EC', fill_type='solid')
        val_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: ContestResults')
    print(f'  Rows 1-5: Title block')
    print(f'  Row 6: spacer')
    print(f'  Row 7: headers')
    print(f'  Rows 8-25: 18 contestants (Score order: random/unsorted)')
    print(f'  Row 26: spacer')
    print(f'  Rows 27-30: summary statistics')


create_initial()
