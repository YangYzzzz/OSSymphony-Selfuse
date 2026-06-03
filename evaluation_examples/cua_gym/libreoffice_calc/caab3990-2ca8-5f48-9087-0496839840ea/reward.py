"""
Reward Script: Extract PDF stats and create LibreOffice Calc table with award rates
Task ID: osworld_multi_apps_pdf_stats_table_009
Domain: libreoffice_calc
Scoring:
  Component 1: Output file teaching_grant_rates.xlsx exists (0.2 pts)
  Component 2: Correct headers in row 1 (0.2 pts)
  Component 3: 6 data rows with correct Year/Applications/Awards values (0.3 pts)
  Component 4: Award Rate (%) computed correctly and formatted to 2 decimal places (0.3 pts)
"""

import os
import sys

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_stats_table_009'

# Expected data from PDFs (2017-2022)
EXPECTED_DATA = {
    2017: {'applications': 342, 'awards': 89,  'award_rate': 26.02},
    2018: {'applications': 401, 'awards': 98,  'award_rate': 24.44},
    2019: {'applications': 387, 'awards': 103, 'award_rate': 26.61},
    2020: {'applications': 312, 'awards': 74,  'award_rate': 23.72},
    2021: {'applications': 428, 'awards': 115, 'award_rate': 26.87},
    2022: {'applications': 456, 'awards': 128, 'award_rate': 28.07},
}

EXPECTED_YEARS = sorted(EXPECTED_DATA.keys())  # [2017, 2018, 2019, 2020, 2021, 2022]


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Component 1: Output file exists (0.2 points)
    # This is the primary gate - without the file there's nothing to verify.
    # On initial_env this file does NOT exist, so this component FAILS there.
    try:
        if not os.path.exists(file_path):
            print(f"FAIL: Component 1 — file not found at {file_path}")
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
        else:
            print(f"PASS: Component 1 — file found at {file_path} (0.2 pts)")
            total_score += 0.2
    except Exception as e:
        print(f"ERROR: Component 1 — cannot check file existence: {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load workbook {file_path}: {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Get active/first sheet
    try:
        ws = wb.active
    except Exception as e:
        print(f"CRITICAL: Cannot access active sheet: {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Correct headers in row 1 (0.2 points)
    # Headers must be: 'Year', 'Applications', 'Awards', 'Award Rate (%)'
    # This FAILS on initial_env (no file) and PASSES on golden_env
    try:
        header_a = ws.cell(row=1, column=1).value
        header_b = ws.cell(row=1, column=2).value
        header_c = ws.cell(row=1, column=3).value
        header_d = ws.cell(row=1, column=4).value

        expected_headers = ['Year', 'Applications', 'Awards', 'Award Rate (%)']
        actual_headers = [header_a, header_b, header_c, header_d]

        # Normalize for comparison (strip whitespace, case-insensitive comparison)
        normalized_actual = [str(h).strip() if h is not None else '' for h in actual_headers]
        normalized_expected = [str(h).strip() for h in expected_headers]

        if normalized_actual == normalized_expected:
            print(f"PASS: Component 2 — headers correct: {actual_headers} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 2 — expected headers {expected_headers}, found {actual_headers}")
    except Exception as e:
        print(f"ERROR: Component 2 — cannot check headers: {e}")

    # Component 3: 6 data rows with correct Year/Applications/Awards values (0.3 points)
    # Rows 2-7 must contain years 2017-2022 with correct application/award counts
    # This FAILS on initial_env (no file) and PASSES on golden_env
    try:
        rows_data = {}
        for row_idx in range(2, 8):  # rows 2-7
            year_val = ws.cell(row=row_idx, column=1).value
            apps_val = ws.cell(row=row_idx, column=2).value
            awards_val = ws.cell(row=row_idx, column=3).value
            if year_val is not None:
                try:
                    rows_data[int(year_val)] = {
                        'applications': apps_val,
                        'awards': awards_val,
                    }
                except (ValueError, TypeError):
                    pass

        # Check all 6 years are present with correct values
        correct_rows = 0
        for year in EXPECTED_YEARS:
            if year not in rows_data:
                print(f"FAIL: Component 3 — year {year} not found in data rows")
                continue
            row = rows_data[year]
            exp = EXPECTED_DATA[year]
            apps_ok = False
            awards_ok = False
            try:
                apps_ok = int(row['applications']) == exp['applications']
            except (ValueError, TypeError):
                pass
            try:
                awards_ok = int(row['awards']) == exp['awards']
            except (ValueError, TypeError):
                pass

            if apps_ok and awards_ok:
                correct_rows += 1
            else:
                print(f"FAIL: Component 3 — year {year}: expected apps={exp['applications']}, awards={exp['awards']}; "
                      f"found apps={row['applications']}, awards={row['awards']}")

        if correct_rows == 6:
            print(f"PASS: Component 3 — all 6 years (2017-2022) have correct Applications and Awards values (0.3 pts)")
            total_score += 0.3
        elif correct_rows >= 3:
            partial = round(0.3 * correct_rows / 6, 2)
            print(f"PARTIAL: Component 3 — {correct_rows}/6 rows correct, partial credit ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — only {correct_rows}/6 rows correct")
    except Exception as e:
        print(f"ERROR: Component 3 — cannot check data rows: {e}")

    # Component 4: Award Rate (%) computed correctly and formatted to 2 decimal places (0.3 points)
    # The Award Rate should be (Awards / Applications) * 100, rounded/formatted to 2 decimal places.
    # The number_format of D column cells should be '0.00' (or similar 2-decimal format).
    # This FAILS on initial_env (no file) and PASSES on golden_env
    try:
        rate_correct_count = 0
        rate_format_count = 0
        tolerance = 0.02  # allow tiny rounding differences

        for row_idx in range(2, 8):  # rows 2-7
            year_val = ws.cell(row=row_idx, column=1).value
            apps_val = ws.cell(row=row_idx, column=2).value
            awards_val = ws.cell(row=row_idx, column=3).value
            rate_val = ws.cell(row=row_idx, column=4).value
            rate_fmt = ws.cell(row=row_idx, column=4).number_format

            if year_val is None:
                continue

            try:
                year = int(year_val)
            except (ValueError, TypeError):
                continue

            if year not in EXPECTED_DATA:
                continue

            exp_rate = EXPECTED_DATA[year]['award_rate']

            # Check rate value
            try:
                actual_rate = float(rate_val)
                if abs(actual_rate - exp_rate) <= tolerance:
                    rate_correct_count += 1
                else:
                    print(f"FAIL: Component 4 (value) — year {year}: expected rate={exp_rate}, found rate={actual_rate}")
            except (ValueError, TypeError):
                print(f"FAIL: Component 4 (value) — year {year}: rate value is not numeric: {rate_val!r}")

            # Check number format is 2-decimal places
            # Valid formats: '0.00', '#,##0.00', '0.00%' (but % would multiply by 100, unlikely)
            # The task says "formatted to 2 decimal places", so look for '0.00' style
            if rate_fmt and ('0.00' in rate_fmt or '0.000' in rate_fmt):
                rate_format_count += 1
            else:
                print(f"FAIL: Component 4 (format) — year {year}: expected 2-decimal format, found {rate_fmt!r}")

        # Award partial points based on correctness
        if rate_correct_count == 6 and rate_format_count == 6:
            print(f"PASS: Component 4 — all 6 award rates correct and formatted to 2 decimal places (0.3 pts)")
            total_score += 0.3
        elif rate_correct_count == 6:
            print(f"PARTIAL: Component 4 — all 6 rates correct but not all formatted to 2 dp "
                  f"({rate_format_count}/6 formatted) (0.2 pts)")
            total_score += 0.2
        elif rate_correct_count >= 3:
            partial = round(0.3 * rate_correct_count / 6, 2)
            print(f"PARTIAL: Component 4 — {rate_correct_count}/6 rates correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — only {rate_correct_count}/6 rates correct")
    except Exception as e:
        print(f"ERROR: Component 4 — cannot check award rates: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Main entry point
file_path = f'{WORKDIR}/teaching_grant_rates.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    try:
        import openpyxl
    except ImportError:
        print("CRITICAL: openpyxl not available on this VM")
        print("REWARD: 0.0")
        sys.exit(1)
    verify_task(file_path)
