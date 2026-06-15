#!/usr/bin/env python3
"""initial_setup.py - Create the initial P&L spreadsheet with headers and input data only."""

import subprocess
import sys

subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
import os
import shlex
import time

output_path = "/home/user/calc_sales_078.xlsx"

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "DealPnL"

# --- Headers ---
headers = ["Deal", "Revenue", "Product Cost %", "Impl Cost", "Discount %",
           "Net Revenue", "Product Cost", "Total Cost", "Gross Profit", "Margin %"]

header_font = Font(bold=True)
header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
header_font_white = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for c, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

# --- Input Data (A2:E5) ---
data = [
    ["D1", 200000, 0.30, 25000, 0.10],
    ["D2", 150000, 0.25, 15000, 0.05],
    ["D3", 500000, 0.35, 50000, 0.15],
    ["D4",  80000, 0.20,  8000, 0.00],
]

for r, row_data in enumerate(data, 2):
    for c, val in enumerate(row_data, 1):
        cell = ws.cell(row=r, column=c, value=val)
        cell.border = border
        if c == 1:
            cell.alignment = Alignment(horizontal="center")
        elif c == 2 or c == 4:
            cell.number_format = '$#,##0'
        elif c == 3 or c == 5:
            cell.number_format = '0%'

# Also add borders for the empty F-J columns in data rows
for r in range(2, 6):
    for c in range(6, 11):
        cell = ws.cell(row=r, column=c)
        cell.border = border

# --- Column widths ---
col_widths = {"A": 10, "B": 14, "C": 16, "D": 14, "E": 14,
              "F": 14, "G": 14, "H": 14, "I": 14, "J": 12}
for col, width in col_widths.items():
    ws.column_dimensions[col].width = width

# Freeze header row
ws.freeze_panes = "A2"

wb.save(output_path)
print(f"Saved initial file to {output_path}")

# Launch LibreOffice Calc
env = os.environ.copy()
env["DISPLAY"] = ":0"
subprocess.Popen(
    shlex.split(f'libreoffice --calc "{output_path}"'),
    stdout=subprocess.DEVNULL,
    stderr=subprocess.DEVNULL,
    env=env,
)
time.sleep(2)
print("LibreOffice Calc launched.")
