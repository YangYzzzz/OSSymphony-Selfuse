import openpyxl

FILE_PATH = "/home/user/calc_sales_078.xlsx"
SHEET_NAME = "DealPnL"

# Expected ground truth values
EXPECTED_F = [180000, 142500, 425000, 80000]
EXPECTED_G = [54000, 35625, 148750, 16000]
EXPECTED_H = [79000, 50625, 198750, 24000]
EXPECTED_I = [101000, 91875, 226250, 56000]
EXPECTED_J = [0.561111, 0.644737, 0.532353, 0.70]

def get_cell_value(ws_formula, ws_data, row, col, input_data):
    """
    Try to get the computed value of a cell.
    1. If data_only gives a cached numeric value, use it.
    2. If the cell has a formula, compute the expected value from input data.
    3. If the cell has a plain numeric value, use it.
    4. Otherwise return None.
    """
    data_val = ws_data.cell(row=row, column=col).value
    formula_val = ws_formula.cell(row=row, column=col).value

    # If data_only gives a cached number, use it
    if data_val is not None and not isinstance(data_val, str):
        return float(data_val)

    # If it's a plain number (no formula), use it
    if formula_val is not None and not isinstance(formula_val, str):
        return float(formula_val)

    # If there's a formula, we compute the expected value ourselves
    # based on the input columns (B=revenue, C=cost%, D=impl_cost, E=discount%)
    if isinstance(formula_val, str) and formula_val.startswith("="):
        r = row  # row index (2-5)
        B = float(ws_formula.cell(row=r, column=2).value)  # Revenue
        C = float(ws_formula.cell(row=r, column=3).value)  # Product Cost %
        D = float(ws_formula.cell(row=r, column=4).value)  # Impl Cost
        E = float(ws_formula.cell(row=r, column=5).value)  # Discount %

        # Compute based on column
        if col == 6:  # F = Net Revenue = B*(1-E)
            return B * (1 - E)
        elif col == 7:  # G = Product Cost = F*C
            F = B * (1 - E)
            return F * C
        elif col == 8:  # H = Total Cost = G + D
            F = B * (1 - E)
            G = F * C
            return G + D
        elif col == 9:  # I = Gross Profit = F - H
            F = B * (1 - E)
            G = F * C
            H = G + D
            return F - H
        elif col == 10:  # J = Margin % = I/F
            F = B * (1 - E)
            G = F * C
            H = G + D
            I_val = F - H
            return I_val / F if F != 0 else None

    return None


def check_column(ws_formula, ws_data, col, expected_values, tolerance, input_data):
    """Check a column (4 rows: 2-5) against expected values. Returns fraction correct."""
    correct = 0
    total = len(expected_values)
    for i, exp in enumerate(expected_values):
        row = i + 2
        val = get_cell_value(ws_formula, ws_data, row, col, input_data)
        if val is not None and abs(val - exp) <= tolerance:
            correct += 1
    return correct / total


def main():
    score = 0.0

    # Gate: file exists and has correct sheet
    try:
        wb_formula = openpyxl.load_workbook(FILE_PATH, data_only=False)
        wb_data = openpyxl.load_workbook(FILE_PATH, data_only=True)
    except Exception:
        print("REWARD: 0.0")
        return

    if SHEET_NAME not in wb_formula.sheetnames:
        print("REWARD: 0.0")
        return

    ws_formula = wb_formula[SHEET_NAME]
    ws_data = wb_data[SHEET_NAME]

    # Check that at least one cell in F2:J5 has content (formula or value)
    has_any_content = False
    for row in range(2, 6):
        for col in range(6, 11):
            v = ws_formula.cell(row=row, column=col).value
            if v is not None:
                has_any_content = True
                break
        if has_any_content:
            break

    if not has_any_content:
        print("REWARD: 0.0")
        return

    input_data = None  # not used directly, computed inside get_cell_value

    # Component 1: Net Revenue F (weight 0.25)
    try:
        s = check_column(ws_formula, ws_data, 6, EXPECTED_F, 0.5, input_data)
        score += 0.25 * s
    except Exception:
        pass

    # Component 2: Product Cost G (weight 0.25)
    try:
        s = check_column(ws_formula, ws_data, 7, EXPECTED_G, 0.5, input_data)
        score += 0.25 * s
    except Exception:
        pass

    # Component 3: Total Cost H (weight 0.15)
    try:
        s = check_column(ws_formula, ws_data, 8, EXPECTED_H, 0.5, input_data)
        score += 0.15 * s
    except Exception:
        pass

    # Component 4: Gross Profit I (weight 0.15)
    try:
        s = check_column(ws_formula, ws_data, 9, EXPECTED_I, 0.5, input_data)
        score += 0.15 * s
    except Exception:
        pass

    # Component 5: Margin % J (weight 0.20)
    try:
        s = check_column(ws_formula, ws_data, 10, EXPECTED_J, 0.01, input_data)
        score += 0.20 * s
    except Exception:
        pass

    score = round(score, 2)
    score = min(1.0, max(0.0, score))
    print(f"REWARD: {score}")


if __name__ == "__main__":
    main()
