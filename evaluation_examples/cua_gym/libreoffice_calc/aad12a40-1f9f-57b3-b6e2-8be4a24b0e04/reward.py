"""
Reward Script: Protect all five sheets and workbook structure with passwords
Task ID: calc_ps_039
Domain: libreoffice_calc
Scoring:
  - Component 1 (0.4): All 5 sheets are protected (sheet=True)
  - Component 2 (0.3): All 5 sheet passwords match hash for 'all2024'
  - Component 3 (0.3): Workbook structure locked with password hash for 'struct2024'
"""

import os
import openpyxl
from openpyxl.worksheet.protection import SheetProtection
from openpyxl.workbook.protection import WorkbookProtection

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_039'

# Pre-compute expected password hashes
_sp = SheetProtection(sheet=True, password='all2024')
EXPECTED_SHEET_HASH = _sp.password  # 'DFFE'

_wp = WorkbookProtection(lockStructure=True, workbookPassword='struct2024')
EXPECTED_WB_HASH = _wp.workbookPassword  # 'C7D6'

EXPECTED_SHEETS = ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4', 'Sheet5']


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: all 5 sheets must exist
    for name in EXPECTED_SHEETS:
        if name not in wb.sheetnames:
            print(f"CRITICAL: Sheet '{name}' not found. Sheets: {wb.sheetnames}")
            print("REWARD: 0.0")
            return 0.0

    # Component 1: All 5 sheets are protected (0.4 points — 0.08 per sheet)
    try:
        sheets_protected = 0
        for name in EXPECTED_SHEETS:
            ws = wb[name]
            if ws.protection.sheet:
                sheets_protected += 1
                print(f"PASS: Component 1 — '{name}' is protected (sheet=True)")
            else:
                print(f"FAIL: Component 1 — '{name}' is NOT protected (sheet={ws.protection.sheet})")
        comp1_score = sheets_protected * 0.08
        if comp1_score > 0:
            total_score += comp1_score
        print(f"Component 1 subtotal: {comp1_score:.2f}/0.40 ({sheets_protected}/5 sheets protected)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: All 5 sheet passwords match 'all2024' hash (0.3 points — 0.06 per sheet)
    try:
        passwords_correct = 0
        for name in EXPECTED_SHEETS:
            ws = wb[name]
            actual_hash = ws.protection.password
            if actual_hash == EXPECTED_SHEET_HASH:
                passwords_correct += 1
                print(f"PASS: Component 2 — '{name}' password hash matches '{EXPECTED_SHEET_HASH}'")
            else:
                print(f"FAIL: Component 2 — '{name}' password hash '{actual_hash}' != expected '{EXPECTED_SHEET_HASH}'")
        comp2_score = passwords_correct * 0.06
        if comp2_score > 0:
            total_score += comp2_score
        print(f"Component 2 subtotal: {comp2_score:.2f}/0.30 ({passwords_correct}/5 correct passwords)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Workbook structure locked with correct password (0.3 points)
    try:
        comp3_score = 0.0
        sec = wb.security
        if sec.lockStructure:
            comp3_score += 0.15
            print(f"PASS: Component 3a — Workbook structure is locked (lockStructure=True)")
        else:
            print(f"FAIL: Component 3a — Workbook structure is NOT locked (lockStructure={sec.lockStructure})")

        if sec.workbookPassword == EXPECTED_WB_HASH:
            comp3_score += 0.15
            print(f"PASS: Component 3b — Workbook password hash matches '{EXPECTED_WB_HASH}'")
        else:
            print(f"FAIL: Component 3b — Workbook password hash '{sec.workbookPassword}' != expected '{EXPECTED_WB_HASH}'")

        if comp3_score > 0:
            total_score += comp3_score
        print(f"Component 3 subtotal: {comp3_score:.2f}/0.30")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
