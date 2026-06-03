"""
Initial Setup: Create workbook with 5 sheets of realistic data, no protection
Task ID: calc_ps_039
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_039'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # --- Sheet1: Employee Directory ---
    ws1 = wb.active
    ws1.title = "Sheet1"
    headers1 = ["Employee ID", "Full Name", "Department", "Position", "Hire Date", "Salary"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border
    data1 = [
        ["EMP001", "Sarah Chen", "Engineering", "Senior Developer", "2021-03-15", 92000],
        ["EMP002", "Marcus Johnson", "Marketing", "Campaign Manager", "2020-07-22", 78000],
        ["EMP003", "Elena Rodriguez", "Finance", "Financial Analyst", "2022-01-10", 85000],
        ["EMP004", "David Kim", "Engineering", "Tech Lead", "2019-11-05", 110000],
        ["EMP005", "Priya Patel", "HR", "HR Specialist", "2023-02-28", 68000],
        ["EMP006", "James O'Brien", "Sales", "Account Executive", "2021-09-14", 75000],
        ["EMP007", "Yuki Tanaka", "Engineering", "QA Engineer", "2022-06-01", 82000],
        ["EMP008", "Maria Santos", "Marketing", "Content Strategist", "2020-12-03", 71000],
        ["EMP009", "Robert Taylor", "Finance", "Controller", "2018-04-18", 125000],
        ["EMP010", "Aisha Mohammed", "Engineering", "DevOps Engineer", "2023-05-20", 95000],
        ["EMP011", "Thomas Weber", "Sales", "Sales Director", "2017-08-12", 130000],
        ["EMP012", "Lisa Nguyen", "HR", "Recruiting Lead", "2021-10-25", 74000],
    ]
    for r, row_data in enumerate(data1, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 20
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 22
    ws1.column_dimensions["E"].width = 14
    ws1.column_dimensions["F"].width = 12

    # --- Sheet2: Quarterly Revenue ---
    ws2 = wb.create_sheet("Sheet2")
    headers2 = ["Region", "Q1 2024", "Q2 2024", "Q3 2024", "Q4 2024", "Annual Total"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border
    data2 = [
        ["North America", 245000, 267000, 298000, 312000, 1122000],
        ["Europe", 189000, 201000, 215000, 228000, 833000],
        ["Asia Pacific", 156000, 178000, 195000, 210000, 739000],
        ["Latin America", 87000, 92000, 101000, 108000, 388000],
        ["Middle East", 45000, 52000, 58000, 63000, 218000],
        ["Africa", 32000, 38000, 42000, 47000, 159000],
        ["Oceania", 28000, 31000, 35000, 39000, 133000],
        ["South Asia", 67000, 74000, 82000, 89000, 312000],
        ["Eastern Europe", 41000, 46000, 51000, 56000, 194000],
        ["Scandinavia", 53000, 58000, 64000, 70000, 245000],
    ]
    for r, row_data in enumerate(data2, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c >= 2:
                cell.number_format = '$#,##0'
    ws2.column_dimensions["A"].width = 18
    for col_letter in ["B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col_letter].width = 14

    # --- Sheet3: Project Tracker ---
    ws3 = wb.create_sheet("Sheet3")
    headers3 = ["Project ID", "Project Name", "Lead", "Status", "Start Date", "Due Date", "Budget"]
    for col, h in enumerate(headers3, 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border
    data3 = [
        ["PRJ-101", "Cloud Migration Phase 2", "David Kim", "In Progress", "2024-01-15", "2024-06-30", 450000],
        ["PRJ-102", "Mobile App Redesign", "Sarah Chen", "Planning", "2024-03-01", "2024-09-15", 280000],
        ["PRJ-103", "Data Warehouse Upgrade", "Robert Taylor", "Completed", "2023-09-01", "2024-02-28", 320000],
        ["PRJ-104", "Customer Portal v3", "Yuki Tanaka", "In Progress", "2024-02-01", "2024-08-31", 195000],
        ["PRJ-105", "AI Chatbot Integration", "Aisha Mohammed", "In Progress", "2024-04-01", "2024-10-31", 175000],
        ["PRJ-106", "ERP System Update", "Elena Rodriguez", "On Hold", "2024-05-15", "2024-12-15", 520000],
        ["PRJ-107", "Security Audit 2024", "James O'Brien", "Planning", "2024-06-01", "2024-08-15", 85000],
        ["PRJ-108", "Brand Refresh Campaign", "Maria Santos", "In Progress", "2024-01-20", "2024-05-30", 120000],
        ["PRJ-109", "Office Relocation IT", "Thomas Weber", "Completed", "2023-11-01", "2024-01-31", 95000],
        ["PRJ-110", "Compliance Training Platform", "Lisa Nguyen", "Planning", "2024-07-01", "2024-11-30", 145000],
    ]
    for r, row_data in enumerate(data3, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 7:
                cell.number_format = '$#,##0'
    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 28
    ws3.column_dimensions["C"].width = 18
    ws3.column_dimensions["D"].width = 14
    ws3.column_dimensions["E"].width = 14
    ws3.column_dimensions["F"].width = 14
    ws3.column_dimensions["G"].width = 12

    # --- Sheet4: Inventory ---
    ws4 = wb.create_sheet("Sheet4")
    headers4 = ["SKU", "Product Name", "Category", "Unit Price", "Quantity", "Reorder Level", "Supplier"]
    for col, h in enumerate(headers4, 1):
        c = ws4.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border
    data4 = [
        ["SKU-2001", "Wireless Mouse Pro", "Electronics", 34.99, 250, 50, "TechSupply Co"],
        ["SKU-2002", "USB-C Hub 7-Port", "Electronics", 49.99, 180, 40, "TechSupply Co"],
        ["SKU-2003", "Ergonomic Keyboard", "Electronics", 79.99, 120, 30, "Peripherals Inc"],
        ["SKU-2004", "Monitor Stand Riser", "Furniture", 45.00, 95, 25, "OfficeFit Ltd"],
        ["SKU-2005", "Desk Organizer Set", "Office Supplies", 22.50, 340, 75, "SupplyChain Plus"],
        ["SKU-2006", "Noise Cancelling Headset", "Electronics", 129.99, 75, 20, "AudioTech Corp"],
        ["SKU-2007", "Webcam HD 1080p", "Electronics", 59.99, 200, 45, "TechSupply Co"],
        ["SKU-2008", "Standing Desk Mat", "Furniture", 38.00, 160, 35, "OfficeFit Ltd"],
        ["SKU-2009", "Whiteboard Markers 12pk", "Office Supplies", 15.99, 420, 100, "SupplyChain Plus"],
        ["SKU-2010", "Laptop Backpack", "Accessories", 64.99, 110, 25, "GearUp Wholesale"],
        ["SKU-2011", "Cable Management Kit", "Accessories", 18.99, 290, 60, "SupplyChain Plus"],
        ["SKU-2012", "Document Scanner", "Electronics", 199.99, 45, 10, "Peripherals Inc"],
    ]
    for r, row_data in enumerate(data4, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws4.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 4:
                cell.number_format = '$#,##0.00'
    ws4.column_dimensions["A"].width = 12
    ws4.column_dimensions["B"].width = 26
    ws4.column_dimensions["C"].width = 18
    ws4.column_dimensions["D"].width = 12
    ws4.column_dimensions["E"].width = 12
    ws4.column_dimensions["F"].width = 14
    ws4.column_dimensions["G"].width = 20

    # --- Sheet5: Meeting Notes ---
    ws5 = wb.create_sheet("Sheet5")
    headers5 = ["Date", "Meeting Title", "Organizer", "Attendees", "Key Decisions", "Action Items"]
    for col, h in enumerate(headers5, 1):
        c = ws5.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border
    data5 = [
        ["2024-01-08", "Q1 Planning Kickoff", "Thomas Weber", "All Dept Heads", "Budget approved for Q1 initiatives", "Submit project proposals by Jan 15"],
        ["2024-01-22", "Engineering Sprint Review", "David Kim", "Engineering Team", "Cloud migration on track", "Complete Phase 2 testing by Feb 1"],
        ["2024-02-05", "Marketing Strategy Session", "Marcus Johnson", "Marketing + Sales", "New campaign targeting SMB segment", "Draft campaign brief by Feb 12"],
        ["2024-02-19", "Finance Review Board", "Robert Taylor", "Finance + Leadership", "Q4 results exceeded targets by 8%", "Prepare investor presentation"],
        ["2024-03-04", "HR Policy Update", "Priya Patel", "HR + Legal", "Remote work policy extended through 2024", "Update employee handbook"],
        ["2024-03-18", "Product Roadmap Review", "Sarah Chen", "Product + Engineering", "AI features prioritized for Q3", "Create technical design docs"],
        ["2024-04-01", "All-Hands Meeting", "Thomas Weber", "All Staff", "Company grew 15% YoY", "Department OKRs due April 8"],
        ["2024-04-15", "Security Incident Debrief", "James O'Brien", "IT + Security", "Phishing attempt blocked successfully", "Roll out new MFA by May 1"],
        ["2024-05-06", "Customer Success Review", "Lisa Nguyen", "CS + Product", "NPS score improved to 72", "Address top 5 feature requests"],
        ["2024-05-20", "Mid-Year Budget Review", "Elena Rodriguez", "Finance + Dept Heads", "Reallocate $50K to AI projects", "Submit revised budgets by May 27"],
    ]
    for r, row_data in enumerate(data5, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws5.cell(row=r, column=c, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws5.column_dimensions["A"].width = 14
    ws5.column_dimensions["B"].width = 28
    ws5.column_dimensions["C"].width = 18
    ws5.column_dimensions["D"].width = 22
    ws5.column_dimensions["E"].width = 36
    ws5.column_dimensions["F"].width = 36

    # No protection applied - initial state must be unprotected
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
