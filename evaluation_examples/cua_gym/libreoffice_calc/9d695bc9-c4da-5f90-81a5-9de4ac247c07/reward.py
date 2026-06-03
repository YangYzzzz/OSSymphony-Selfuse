"""
Reward Script: Calculate monthly mortgage payment using PMT function
Task ID: calc_fmb_pmt_loan_029
Domain: libreoffice_calc
Scoring:
  Component 1: B6 contains a PMT formula (0.5 pts)
  Component 2: PMT formula uses correct cell references B3/B5, B4*B5, -B2 (0.5 pts)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmb_pmt_loan_029'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    The task requires placing =PMT(B3/B5,B4*B5,-B2) in cell B6 of the
    'Mortgage Calculator' sheet. The initial file has B6 empty.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Mortgage Calculator' sheet must exist
    if 'Mortgage Calculator' not in wb.sheetnames:
        print("FAIL: 'Mortgage Calculator' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Mortgage Calculator']

    # Component 1: B6 contains a PMT formula (0.5 points)
    # This FAILS on initial (B6 is empty) → PASSES on golden (B6 has PMT formula)
    try:
        b6_value = ws['B6'].value
        if b6_value is not None and isinstance(b6_value, str) and 'PMT' in b6_value.upper():
            print(f"PASS: Component 1 — B6 contains a PMT formula: {repr(b6_value)} (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 — B6 expected to contain a PMT formula, found: {repr(b6_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — Cannot read B6: {e}")

    # Component 2: PMT formula uses correct cell references (0.5 points)
    # Checks: rate=B3/B5, nper=B4*B5, pv=-B2
    # Canonical form: =PMT(B3/B5,B4*B5,-B2)
    # This FAILS on initial (B6 is empty) → PASSES on golden (correct PMT formula)
    try:
        b6_value = ws['B6'].value
        if b6_value is not None and isinstance(b6_value, str):
            # Normalize: remove spaces, uppercase for comparison
            formula_normalized = b6_value.upper().replace(' ', '')
            # Check the formula contains the required argument structure
            # Accept both =PMT(B3/B5,B4*B5,-B2) and PMT with negation on pv
            # The formula must reference: rate=B3/B5, nper=B4*B5, pv using -B2 (or B2 negated)
            has_rate = bool(re.search(r'PMT\(B3/B5', formula_normalized))
            has_nper = bool(re.search(r'B3/B5,B4\*B5', formula_normalized))
            has_pv = bool(re.search(r'B4\*B5,-B2\)', formula_normalized))
            if has_rate and has_nper and has_pv:
                print(f"PASS: Component 2 — PMT formula uses correct references (rate=B3/B5, nper=B4*B5, pv=-B2): {repr(b6_value)} (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — PMT formula does not use correct references.")
                print(f"  Expected: =PMT(B3/B5,B4*B5,-B2)")
                print(f"  Found: {repr(b6_value)}")
                print(f"  has_rate(B3/B5)={has_rate}, has_nper(B4*B5)={has_nper}, has_pv(-B2)={has_pv}")
        else:
            print(f"FAIL: Component 2 — B6 is empty or not a formula string, found: {repr(b6_value)}")
    except Exception as e:
        print(f"ERROR: Component 2 — Cannot verify formula references: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
