"""
Initial Setup: Mortgage calculator spreadsheet with loan parameters but empty payment cell.
Task ID: calc_fmb_pmt_loan_029
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

WORKDIR = '/home/user'
TASK_ID = 'calc_fmb_pmt_loan_029'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Mortgage Calculator'

    # Column widths for readability
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18

    # Title row
    ws['A1'] = 'Home Loan Mortgage Calculator'
    ws['A1'].font = Font(name='Calibri', size=14, bold=True)
    ws.merge_cells('A1:B1')
    ws['A1'].alignment = Alignment(horizontal='center')

    # Header style for labels
    label_font = Font(name='Calibri', size=11)
    value_font = Font(name='Calibri', size=11)

    # Row 2: Loan Amount
    ws['A2'] = 'Loan Amount'
    ws['A2'].font = label_font
    ws['B2'] = 285000
    ws['B2'].font = value_font
    ws['B2'].number_format = '$#,##0.00'

    # Row 3: Annual Interest Rate
    ws['A3'] = 'Annual Interest Rate'
    ws['A3'].font = label_font
    ws['B3'] = 0.045
    ws['B3'].font = value_font
    ws['B3'].number_format = '0.00%'

    # Row 4: Loan Term (Years)
    ws['A4'] = 'Loan Term (Years)'
    ws['A4'].font = label_font
    ws['B4'] = 30
    ws['B4'].font = value_font
    ws['B4'].number_format = '0'

    # Row 5: Payments per Year
    ws['A5'] = 'Payments per Year'
    ws['A5'].font = label_font
    ws['B5'] = 12
    ws['B5'].font = value_font
    ws['B5'].number_format = '0'

    # Row 6: Monthly Payment — B6 is intentionally EMPTY (the target cell)
    ws['A6'] = 'Monthly Payment'
    ws['A6'].font = Font(name='Calibri', size=11, bold=True)
    # B6 is left empty — user must add the PMT formula

    # Light background for label column
    label_fill = PatternFill(start_color='FFE2EFDA', end_color='FFE2EFDA', fill_type='solid')
    for row in range(2, 7):
        ws[f'A{row}'].fill = label_fill

    # Border around the data range A1:B6
    thin = Side(style='thin', color='000000')
    for row in range(1, 7):
        for col in ['A', 'B']:
            ws[f'{col}{row}'].border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
