"""
Reward Script: International currency expense report in LibreOffice Calc
Task ID: calc_grs_075
Domain: libreoffice_calc
Scoring:
  Component 1 (0.30): G column has =E*F formulas for Amount in USD (G2:G24)
  Component 2 (0.15): Total reimbursable SUM formula in G25
  Component 3 (0.15): Conditional formatting rule on Expense Log (>$500 flag)
  Component 4 (0.25): Summary sheet has currency summary table with formulas
  Component 5 (0.15): Summary sheet has a chart (category breakdown pie chart)
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_075'


def persist_app_state(domain):
    """Save any unsaved LibreOffice state before verification."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        import time
        time.sleep(1.0)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Expense Log' sheet must exist
    if 'Expense Log' not in wb.sheetnames:
        print("CRITICAL: 'Expense Log' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_log = wb['Expense Log']

    # ------------------------------------------------------------------
    # Component 1: G column formulas (Amount in USD) — 0.30 points
    # Initial has NO values in G column. Golden has =E*F formulas in G2:G24.
    # ------------------------------------------------------------------
    try:
        formula_count = 0
        valid_formula_count = 0
        for row_num in range(2, 25):  # G2:G24 = 23 cells
            cell_val = ws_log.cell(row=row_num, column=7).value  # Column G
            if cell_val is not None:
                formula_count += 1
                val_str = str(cell_val).upper().replace(" ", "")
                # Accept formulas like =E2*F2 or =F2*E2 or similar multiplication
                if re.match(r'^=.*E\d+.*\*.*F\d+|^=.*F\d+.*\*.*E\d+', val_str):
                    valid_formula_count += 1

        if valid_formula_count >= 20:
            # Most or all formulas are correct multiplication formulas
            print(f"PASS: Component 1 — {valid_formula_count}/23 G column cells have E*F formulas (0.30 pts)")
            total_score += 0.30
        elif valid_formula_count >= 10:
            partial = 0.15
            print(f"PARTIAL: Component 1 — {valid_formula_count}/23 G column formulas found ({partial} pts)")
            total_score += partial
        elif formula_count >= 10:
            # Has values but maybe not formulas (could be computed values)
            print(f"PARTIAL: Component 1 — {formula_count}/23 G column cells populated, {valid_formula_count} with formulas (0.10 pts)")
            if formula_count >= 10:
                total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Only {formula_count}/23 G column cells populated, {valid_formula_count} with formulas")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ------------------------------------------------------------------
    # Component 2: Total reimbursable amount SUM formula — 0.15 points
    # Initial has no row 25. Golden has G25=SUM(G2:G24) with bold formatting.
    # ------------------------------------------------------------------
    try:
        # Check for a SUM formula in the G column, anywhere from row 25 onwards
        sum_row = None
        for row_num in range(25, 35):  # Search a range in case row shifted
            cell_val = ws_log.cell(row=row_num, column=7).value
            if cell_val is not None and isinstance(cell_val, str):
                val_upper = cell_val.upper().replace(" ", "")
                if '=SUM(' in val_upper and 'G' in val_upper:
                    sum_row = row_num
                    break

        if sum_row is not None:
            print(f"PASS: Component 2 — SUM formula found in G{sum_row}: {ws_log.cell(row=sum_row, column=7).value} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 2 — No SUM formula found in G column for total reimbursable amount")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ------------------------------------------------------------------
    # Component 3: Conditional formatting on Expense Log — 0.15 points
    # Initial has 0 conditional formatting rules. Golden has 1 rule.
    # Task requires flagging expenses over $500 in red.
    # ------------------------------------------------------------------
    try:
        cf_rules = list(ws_log.conditional_formatting)
        if len(cf_rules) > 0:
            print(f"PASS: Component 3 — {len(cf_rules)} conditional formatting rule(s) found on Expense Log (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 3 — No conditional formatting rules on Expense Log (expected >$500 flag)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ------------------------------------------------------------------
    # Component 4: Summary sheet currency summary table — 0.25 points
    # Initial Summary has only A1="Currency Summary" with no table data.
    # Golden has currency summary (A3:C9) with SUMPRODUCT formulas and
    # category breakdown (A11:B17).
    # ------------------------------------------------------------------
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            # Check for currency summary table: currencies in A4:A8
            currencies_expected = {'USD', 'GBP', 'JPY', 'SGD', 'AUD'}
            currencies_found = set()
            for row_num in range(3, 15):
                val = ws_sum.cell(row=row_num, column=1).value
                if val and str(val).strip().upper() in currencies_expected:
                    currencies_found.add(str(val).strip().upper())

            # Check for formulas in C column (USD equivalent totals)
            formula_cells_found = 0
            for row_num in range(3, 15):
                c_val = ws_sum.cell(row=row_num, column=3).value
                if c_val is not None and isinstance(c_val, str) and c_val.startswith('='):
                    formula_cells_found += 1

            # Also check for category breakdown section
            category_keywords = {'transportation', 'accommodation', 'meals', 'conference', 'supplies'}
            categories_found = set()
            for row_num in range(10, 25):
                val = ws_sum.cell(row=row_num, column=1).value
                if val and str(val).strip().lower() in category_keywords:
                    categories_found.add(str(val).strip().lower())

            score_4 = 0.0
            if len(currencies_found) >= 4:
                score_4 += 0.10
                print(f"  SUB-PASS: Currency summary has {len(currencies_found)}/5 currencies")
            else:
                print(f"  SUB-FAIL: Currency summary has {len(currencies_found)}/5 currencies: {currencies_found}")

            if formula_cells_found >= 3:
                score_4 += 0.08
                print(f"  SUB-PASS: {formula_cells_found} formula cells in summary C column")
            else:
                print(f"  SUB-FAIL: Only {formula_cells_found} formula cells in summary C column")

            if len(categories_found) >= 3:
                score_4 += 0.07
                print(f"  SUB-PASS: Category breakdown has {len(categories_found)}/5 categories")
            else:
                print(f"  SUB-FAIL: Category breakdown has {len(categories_found)}/5 categories: {categories_found}")

            if score_4 > 0:
                print(f"PASS: Component 4 — Summary table score: {score_4} pts")
                total_score += score_4
            else:
                print(f"FAIL: Component 4 — Summary sheet lacks currency summary table and category breakdown")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ------------------------------------------------------------------
    # Component 5: Chart on Summary sheet — 0.15 points
    # Initial Summary has 0 charts. Golden has 1 chart (category pie chart).
    # ------------------------------------------------------------------
    try:
        if 'Summary' not in wb.sheetnames:
            print("FAIL: Component 5 — 'Summary' sheet not found")
        else:
            ws_sum = wb['Summary']
            chart_count = len(ws_sum._charts)
            if chart_count >= 1:
                print(f"PASS: Component 5 — {chart_count} chart(s) found on Summary sheet (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 — No charts on Summary sheet (expected category breakdown pie chart)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
