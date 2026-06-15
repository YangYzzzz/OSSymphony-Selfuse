"""
Initial Setup: International currency expense report for business trip
Task ID: calc_grs_075
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_075'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ========== Sheet 1: Expense Log ==========
    ws1 = wb.active
    ws1.title = "Expense Log"

    headers = [
        "Date", "Description", "Category", "Currency",
        "Amount (Foreign)", "Exchange Rate", "Amount (USD)", "Receipt #"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Realistic expense data from a 2-week business trip
    # Columns: Date, Description, Category, Currency, Amount(Foreign), Exchange Rate, Amount(USD), Receipt#
    expenses = [
        ["2026-03-10", "Heathrow Express to Paddington", "Transportation", "GBP", 25.00, 1.27, None, "REC-0310-001"],
        ["2026-03-10", "Premier Inn London City", "Accommodation", "GBP", 189.00, 1.27, None, "REC-0310-002"],
        ["2026-03-11", "Uber to ExCeL London Centre", "Transportation", "GBP", 32.50, 1.27, None, "REC-0311-001"],
        ["2026-03-11", "Conference Registration - London Tech Summit", "Conference", "GBP", 450.00, 1.27, None, "REC-0311-002"],
        ["2026-03-11", "Client dinner at The Ivy", "Meals", "GBP", 187.40, 1.27, None, "REC-0311-003"],
        ["2026-03-12", "British Airways LHR to NRT", "Transportation", "USD", 1245.00, 1.00, None, "REC-0312-001"],
        ["2026-03-13", "Hotel Gracery Shinjuku (3 nights)", "Accommodation", "JPY", 54000, 0.0067, None, "REC-0313-001"],
        ["2026-03-13", "Suica card top-up", "Transportation", "JPY", 5000, 0.0067, None, "REC-0313-002"],
        ["2026-03-14", "Partner meeting lunch - Roppongi", "Meals", "JPY", 8500, 0.0067, None, "REC-0314-001"],
        ["2026-03-14", "Tokyo conference venue fee", "Conference", "JPY", 35000, 0.0067, None, "REC-0314-002"],
        ["2026-03-15", "Shinkansen Tokyo to Osaka day trip", "Transportation", "JPY", 13620, 0.0067, None, "REC-0315-001"],
        ["2026-03-16", "ANA flight NRT to SIN", "Transportation", "SGD", 580.00, 0.75, None, "REC-0316-001"],
        ["2026-03-17", "Marina Bay Sands (2 nights)", "Accommodation", "SGD", 890.00, 0.75, None, "REC-0317-001"],
        ["2026-03-17", "Grab to Singapore Expo", "Transportation", "SGD", 22.50, 0.75, None, "REC-0317-002"],
        ["2026-03-18", "APAC Partner Summit registration", "Conference", "SGD", 320.00, 0.75, None, "REC-0318-001"],
        ["2026-03-18", "Team dinner at Jumbo Seafood", "Meals", "SGD", 245.60, 0.75, None, "REC-0318-002"],
        ["2026-03-19", "Singapore Airlines SIN to SYD", "Transportation", "AUD", 720.00, 0.65, None, "REC-0319-001"],
        ["2026-03-20", "Hilton Sydney (3 nights)", "Accommodation", "AUD", 1050.00, 0.65, None, "REC-0320-001"],
        ["2026-03-20", "Opal card for trains", "Transportation", "AUD", 40.00, 0.65, None, "REC-0320-002"],
        ["2026-03-21", "Sydney Tech Conference pass", "Conference", "AUD", 395.00, 0.65, None, "REC-0321-001"],
        ["2026-03-21", "Business lunch at Quay Restaurant", "Meals", "AUD", 310.00, 0.65, None, "REC-0321-002"],
        ["2026-03-22", "Office supplies - presentation materials", "Supplies", "AUD", 85.50, 0.65, None, "REC-0322-001"],
        ["2026-03-23", "Qantas SYD to LAX", "Transportation", "USD", 1580.00, 1.00, None, "REC-0323-001"],
    ]

    for r, row_data in enumerate(expenses, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 1:  # Date
                cell.number_format = 'yyyy-mm-dd'
                cell.alignment = Alignment(horizontal="center")
            elif c == 5:  # Amount (Foreign)
                cell.number_format = '#,##0.00'
            elif c == 6:  # Exchange Rate
                cell.number_format = '0.0000'
            elif c == 7:  # Amount (USD) - intentionally blank
                pass
            elif c == 8:  # Receipt #
                cell.alignment = Alignment(horizontal="center")

    # Category dropdown validation
    dv_category = DataValidation(
        type="list",
        formula1='"Transportation,Accommodation,Meals,Conference,Supplies,Entertainment,Communication"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_category.error = "Please select a valid category"
    dv_category.errorTitle = "Invalid Category"
    dv_category.add("C2:C100")
    ws1.add_data_validation(dv_category)

    # Currency dropdown validation
    dv_currency = DataValidation(
        type="list",
        formula1='"USD,EUR,GBP,JPY,AUD,CAD,SGD"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_currency.error = "Please select a valid currency"
    dv_currency.errorTitle = "Invalid Currency"
    dv_currency.add("D2:D100")
    ws1.add_data_validation(dv_currency)

    # Set column widths for readability
    ws1.column_dimensions["A"].width = 14
    ws1.column_dimensions["B"].width = 42
    ws1.column_dimensions["C"].width = 18
    ws1.column_dimensions["D"].width = 12
    ws1.column_dimensions["E"].width = 20
    ws1.column_dimensions["F"].width = 16
    ws1.column_dimensions["G"].width = 18
    ws1.column_dimensions["H"].width = 16

    # Freeze header row
    ws1.freeze_panes = "A2"

    # ========== Sheet 2: Exchange Rates ==========
    ws2 = wb.create_sheet("Exchange Rates")

    rate_headers = ["Currency Code", "Currency Name", "Rate to USD", "Last Updated"]
    rate_header_fill = PatternFill(start_color="FF375623", end_color="FF375623", fill_type="solid")
    rate_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

    for col, h in enumerate(rate_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = rate_header_font
        cell.fill = rate_header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    rate_data = [
        ["USD", "US Dollar", 1.0000, "2026-03-23"],
        ["EUR", "Euro", 1.0850, "2026-03-23"],
        ["GBP", "British Pound Sterling", 1.2700, "2026-03-23"],
        ["JPY", "Japanese Yen", 0.0067, "2026-03-23"],
        ["AUD", "Australian Dollar", 0.6500, "2026-03-23"],
        ["CAD", "Canadian Dollar", 0.7400, "2026-03-23"],
        ["SGD", "Singapore Dollar", 0.7500, "2026-03-23"],
    ]

    for r, row_data in enumerate(rate_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.border = thin_border
            if c == 3:
                cell.number_format = '0.0000'
            elif c == 4:
                cell.number_format = 'yyyy-mm-dd'
            if c == 1:
                cell.alignment = Alignment(horizontal="center")

    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 28
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16

    # ========== Sheet 3: Summary (empty placeholder) ==========
    ws3 = wb.create_sheet("Summary")
    ws3.cell(row=1, column=1, value="Currency Summary")
    ws3.cell(row=1, column=1).font = Font(name="Calibri", size=14, bold=True)
    # Leave rest empty - agent needs to build this

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
