"""
Initial Setup: Sales Quota Progress Chart
Task ID: calc_sales_quota_progress_chart_034
Domain: libreoffice_calc

Creates the initial state for the quota visualization task.
Sheet 'QuotaViz' with 11 reps and quota/actual data.
No charts present — chart creation is the agent's task.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_quota_progress_chart_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: QuotaViz ---
    ws = wb.active
    ws.title = 'QuotaViz'

    # Headers A1:D1
    headers = ['Rep Name', 'Quota', 'Actual Sales', 'Attainment %']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, name='Calibri', size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Sales rep data: 11 reps
    # Quota range: $500,000 to $1,200,000
    # Actual range: $320,000 to $1,450,000
    rep_data = [
        ('Sarah Chen',       1200000,  1350000),
        ('Marcus Johnson',    950000,  1100000),
        ('Priya Patel',      1050000,   870000),
        ('Derek Williams',    800000,   980000),
        ('Lisa Nguyen',       700000,   320000),
        ('Carlos Rivera',    1100000,  1450000),
        ('Rachel Thompson',   600000,   615000),
        ('James O\'Brien',    850000,   790000),
        ('Amira Hassan',      500000,   530000),
        ('Tyler Brooks',      900000,   720000),
        ('Yuna Kim',         1000000,   995000),
    ]

    for r, (name, quota, actual) in enumerate(rep_data, 2):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=quota)
        ws.cell(row=r, column=3, value=actual)
        # Attainment % formula: =C/B (as percentage)
        ws.cell(row=r, column=4, value=f'=C{r}/B{r}')
        # Format quota and actual as currency
        ws.cell(row=r, column=2).number_format = '$#,##0'
        ws.cell(row=r, column=3).number_format = '$#,##0'
        # Format attainment as percentage
        ws.cell(row=r, column=4).number_format = '0.0%'

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Sheet: QuotaViz with 11 reps (rows 2-12)')
    print(f'Columns: Rep Name, Quota, Actual Sales, Attainment %')
    print(f'No charts present (chart creation is the agent task)')


create_initial()
