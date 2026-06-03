"""
Initial Setup: Unmerge cells, fill down categories, and apply AutoFilter
Task ID: calc_tbl_022
Domain: libreoffice_calc

Creates a product inventory spreadsheet with merged category cells in column A.
AutoFilter cannot be applied because of the merged cells.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_022'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Product Inventory"

    # --- Headers ---
    headers = ['Category', 'Product', 'SKU', 'Price', 'Stock', 'Supplier']
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data by category ---
    # Category: Electronics (rows 2-5)
    electronics = [
        ['Wireless Earbuds Pro', 'WEP-4521', 79.99, 234, 'TechVault Inc.'],
        ['USB-C Hub 7-in-1', 'UCH-7100', 45.50, 156, 'ConnectPro Ltd.'],
        ['Mechanical Keyboard RGB', 'MKR-8830', 129.99, 87, 'KeyMaster Corp.'],
        ['Portable SSD 1TB', 'PSD-1024', 94.95, 312, 'DataStream LLC'],
    ]

    # Category: Clothing (rows 6-9)
    clothing = [
        ['Merino Wool Sweater', 'MWS-2200', 89.00, 145, 'Alpine Threads'],
        ['Slim Fit Chinos', 'SFC-3340', 54.99, 278, 'UrbanWear Co.'],
        ['Waterproof Jacket', 'WPJ-5500', 149.95, 63, 'StormShield Ltd.'],
        ['Running Shoes Elite', 'RSE-7780', 119.00, 192, 'PaceFit Inc.'],
    ]

    # Category: Home & Garden (rows 10-13)
    home_garden = [
        ['Ceramic Plant Pot Set', 'CPP-1100', 34.99, 421, 'GreenSpace Supplies'],
        ['LED Desk Lamp', 'LDL-4400', 42.50, 189, 'BrightLife Co.'],
        ['Bamboo Cutting Board', 'BCB-6600', 28.75, 356, 'EcoKitchen Ltd.'],
        ['Smart Thermostat', 'STH-9900', 199.00, 74, 'HomeLogic Inc.'],
    ]

    # Category: Sports & Outdoors (rows 14-17)
    sports = [
        ['Yoga Mat Premium', 'YMP-3300', 39.99, 267, 'ZenFit Gear'],
        ['Camping Headlamp', 'CHL-5550', 24.95, 443, 'TrailBlaze Co.'],
        ['Resistance Band Set', 'RBS-7700', 19.99, 581, 'FlexForce Ltd.'],
        ['Insulated Water Bottle', 'IWB-8800', 32.50, 398, 'HydroTrek Inc.'],
    ]

    categories = [
        ('Electronics', electronics, 2),
        ('Clothing', clothing, 6),
        ('Home & Garden', home_garden, 10),
        ('Sports & Outdoors', sports, 14),
    ]

    data_align = Alignment(vertical="center")
    price_fmt = '$#,##0.00'

    for cat_name, items, start_row in categories:
        for i, item in enumerate(items):
            row = start_row + i
            # Column A (category) - will be merged later, only write to first row
            if i == 0:
                ws.cell(row=row, column=1, value=cat_name)
            # Columns B-F
            for c, val in enumerate(item, 2):
                cell = ws.cell(row=row, column=c, value=val)
                cell.alignment = data_align
                cell.border = thin_border
                if c == 4:  # Price column
                    cell.number_format = price_fmt

        # Apply border and alignment to column A cells before merging
        for i in range(4):
            row = start_row + i
            cell = ws.cell(row=row, column=1)
            cell.border = thin_border

        # Merge category cells in column A
        merge_range = f'A{start_row}:A{start_row + 3}'
        ws.merge_cells(merge_range)
        # Style the merged cell (top-left)
        top_cell = ws.cell(row=start_row, column=1)
        top_cell.font = Font(name="Calibri", size=11, bold=True)
        top_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 22

    # --- Row height for header ---
    ws.row_dimensions[1].height = 22

    # NO AutoFilter - that's the task!

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
