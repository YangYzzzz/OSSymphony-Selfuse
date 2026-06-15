"""
Reward Script: Add a hyperlink to cell A1 that opens https://www.acmecorp.example.com,
               displaying the text 'Visit ACME Website'.
Task ID: calc_cop_hyperlink_001
Domain: libreoffice_calc
Scoring:
  - Component 1: Cell A1 displays text 'Visit ACME Website'         (0.3 pts)
  - Component 2: Cell A1 has hyperlink to https://www.acmecorp.example.com  (0.5 pts)
  - Component 3: Cell A1 is formatted as a hyperlink (underlined)   (0.2 pts)
  Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_cop_hyperlink_001'
SHEET_NAME = 'Dashboard'
TARGET_CELL = 'A1'
EXPECTED_TEXT = 'Visit ACME Website'
EXPECTED_URL = 'https://www.acmecorp.example.com'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: sheet 'Dashboard' must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # Component 1: Cell A1 displays the text 'Visit ACME Website' (0.3 points)
    # This FAILS on initial (A1 is empty) and PASSES on golden
    try:
        cell_value = ws[TARGET_CELL].value
        if cell_value is not None and str(cell_value).strip() == EXPECTED_TEXT:
            print(f"PASS: Component 1 — A1 displays text '{cell_value}' (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Expected A1 text '{EXPECTED_TEXT}', found: {repr(cell_value)}")
    except Exception as e:
        print(f"ERROR: Component 1 — could not read A1 value: {e}")

    # Component 2: Cell A1 has a hyperlink targeting the expected URL (0.5 points)
    # This FAILS on initial (no hyperlink exists) and PASSES on golden
    try:
        hyperlink = ws[TARGET_CELL].hyperlink
        if hyperlink is not None:
            target = hyperlink.target
            if target is not None and target.strip() == EXPECTED_URL:
                print(f"PASS: Component 2 — A1 hyperlink target is '{target}' (0.5 pts)")
                total_score += 0.5
            else:
                print(f"FAIL: Component 2 — Expected hyperlink target '{EXPECTED_URL}', found: {repr(target)}")
        else:
            print(f"FAIL: Component 2 — A1 has no hyperlink")
    except Exception as e:
        print(f"ERROR: Component 2 — could not check A1 hyperlink: {e}")

    # Component 3: Cell A1 is formatted with underline (hyperlink formatting) (0.2 points)
    # This FAILS on initial (no underline) and PASSES on golden (single underline)
    try:
        underline = ws[TARGET_CELL].font.underline
        if underline is not None and underline != 'none' and underline is not False:
            print(f"PASS: Component 3 — A1 has underline formatting: '{underline}' (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 — Expected A1 to have underline, found underline={repr(underline)}")
    except Exception as e:
        print(f"ERROR: Component 3 — could not check A1 font underline: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
