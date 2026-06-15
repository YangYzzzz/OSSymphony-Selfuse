"""
Initial Setup: Dashboard spreadsheet for hyperlink task
Task ID: calc_cop_hyperlink_001
Domain: libreoffice_calc

Creates a 'Dashboard' sheet with KPI data.
Cell A1 is EMPTY — the task requires adding a hyperlink there.
No hyperlinks exist anywhere in the sheet.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_cop_hyperlink_001'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Dashboard ---
    ws = wb.active
    ws.title = 'Dashboard'

    # Leave A1 empty — hyperlink will be added here by the agent
    # Row 1: Section header (B1 onwards)
    ws['B1'] = 'ACME Corporation — Q1 2025 Performance Dashboard'
    ws['B1'].font = Font(name='Calibri', size=14, bold=True)
    ws['B1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28

    # Spacer row 2 is blank

    # KPI section headers (row 3)
    kpi_headers = ['Metric', 'Target', 'Actual', 'Variance', 'Status']
    for col, h in enumerate(kpi_headers, 2):  # B–F
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF2E4057', end_color='FF2E4057', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[3].height = 22

    # KPI data rows (rows 4–13)
    kpi_data = [
        ['Total Revenue ($)',       4500000, 4823150, 323150,   'Above Target'],
        ['Gross Profit Margin (%)',      38,    41.2,     3.2,  'Above Target'],
        ['Operating Expenses ($)',  1200000, 1183400,  -16600,  'On Track'],
        ['New Customers',               850,     912,      62,  'Above Target'],
        ['Customer Retention (%)',       88,    85.7,    -2.3,  'Below Target'],
        ['Net Promoter Score',           55,      59,       4,  'Above Target'],
        ['Support Tickets Resolved',    920,     897,     -23,  'Below Target'],
        ['Product Defect Rate (%)',     1.5,     1.1,    -0.4,  'Above Target'],
        ['Employee Satisfaction (%)',    78,    80.5,     2.5,  'On Track'],
        ['Market Share (%)',            22,    23.8,     1.8,  'Above Target'],
    ]

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r, (metric, target, actual, variance, status) in enumerate(kpi_data, 4):
        row_fill_color = 'FFFAFAFA' if r % 2 == 0 else 'FFFFFFFF'
        row_fill = PatternFill(start_color=row_fill_color, end_color=row_fill_color, fill_type='solid')

        # Col B: metric name
        cell_b = ws.cell(row=r, column=2, value=metric)
        cell_b.font = Font(name='Calibri', size=10)
        cell_b.fill = row_fill
        cell_b.border = border
        cell_b.alignment = Alignment(horizontal='left', vertical='center')

        # Col C: target
        cell_c = ws.cell(row=r, column=3, value=target)
        cell_c.font = Font(name='Calibri', size=10)
        cell_c.fill = row_fill
        cell_c.border = border
        cell_c.alignment = Alignment(horizontal='right', vertical='center')

        # Col D: actual
        cell_d = ws.cell(row=r, column=4, value=actual)
        cell_d.font = Font(name='Calibri', size=10)
        cell_d.fill = row_fill
        cell_d.border = border
        cell_d.alignment = Alignment(horizontal='right', vertical='center')

        # Col E: variance (color-coded)
        cell_e = ws.cell(row=r, column=5, value=variance)
        var_color = '00008000' if variance >= 0 else '00CC0000'
        cell_e.font = Font(name='Calibri', size=10, color=var_color)
        cell_e.fill = row_fill
        cell_e.border = border
        cell_e.alignment = Alignment(horizontal='right', vertical='center')

        # Col F: status
        cell_f = ws.cell(row=r, column=6, value=status)
        if status == 'Above Target':
            status_color = 'FF00AA44'
        elif status == 'Below Target':
            status_color = 'FFCC2200'
        else:
            status_color = 'FF886600'
        cell_f.font = Font(name='Calibri', size=10, bold=True, color=status_color[2:])
        cell_f.fill = row_fill
        cell_f.border = border
        cell_f.alignment = Alignment(horizontal='center', vertical='center')

    # Summary row (row 14)
    ws.cell(row=14, column=2, value='Summary').font = Font(name='Calibri', size=10, bold=True)
    ws.cell(row=14, column=4, value='=SUM(D4:D13)').number_format = '#,##0'
    ws.cell(row=14, column=5, value='=SUM(E4:E13)').number_format = '#,##0'

    # Column widths
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 32
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 18

    # Freeze panes below header rows
    ws.freeze_panes = 'B4'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')


create_initial()
