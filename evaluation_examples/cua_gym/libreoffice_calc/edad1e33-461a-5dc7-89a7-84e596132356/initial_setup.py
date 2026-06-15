"""
Initial Setup: Resize all columns A-H to optimal width in a contact database spreadsheet.
Task ID: calc_gfl_057
Domain: libreoffice_calc

Creates a Contacts spreadsheet with 44 records where columns have deliberately
wrong widths (some too narrow, some too wide) so the agent must auto-fit them.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_057'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contacts'

    # Headers
    headers = ['First Name', 'Last Name', 'Company', 'Job Title', 'Email', 'Phone', 'City', 'Country']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # 44 contact records with realistic data
    contacts = [
        ['Sarah', 'Chen', 'Nextera Technologies', 'Senior Software Engineer', 'sarah.chen@nextera-tech.com', '+1-415-555-0142', 'San Francisco', 'United States'],
        ['Marcus', 'Johnson', 'Pinnacle Financial Group', 'Portfolio Manager', 'marcus.johnson@pinnacle-fg.com', '+1-212-555-0198', 'New York', 'United States'],
        ['Elena', 'Kowalski', 'Baltic Shipping Solutions', 'Operations Director', 'elena.kowalski@balticshipping.pl', '+48-22-555-0173', 'Warsaw', 'Poland'],
        ['Raj', 'Patel', 'CloudBridge Analytics', 'Data Scientist', 'raj.patel@cloudbridge.io', '+91-80-5555-0234', 'Bangalore', 'India'],
        ['Amara', 'Okafor', 'GreenLeaf Organics', 'Marketing Manager', 'amara.okafor@greenleaf-organics.ng', '+234-1-555-0187', 'Lagos', 'Nigeria'],
        ['Thomas', 'Mueller', 'Precision AutoWerks', 'Quality Assurance Lead', 'thomas.mueller@precision-aw.de', '+49-89-555-0156', 'Munich', 'Germany'],
        ['Yuki', 'Tanaka', 'Sakura Digital Media', 'Creative Director', 'yuki.tanaka@sakura-dm.co.jp', '+81-3-5555-0291', 'Tokyo', 'Japan'],
        ['Isabella', 'Rossi', 'Mediterraneo Hospitality', 'General Manager', 'isabella.rossi@mediterraneo.it', '+39-06-555-0124', 'Rome', 'Italy'],
        ['David', 'Oconnell', 'Atlantic Research Group', 'Principal Investigator', 'david.oconnell@atlantic-rg.ie', '+353-1-555-0168', 'Dublin', 'Ireland'],
        ['Li', 'Wei', 'Dragon Gate Imports', 'Supply Chain Manager', 'li.wei@dragongate-imports.cn', '+86-21-5555-0345', 'Shanghai', 'China'],
        ['Fatima', 'Al-Hassan', 'Oasis Development Corp', 'Project Coordinator', 'fatima.alhassan@oasis-dev.ae', '+971-4-555-0213', 'Dubai', 'United Arab Emirates'],
        ['Carlos', 'Mendez', 'SolTech Renewables', 'Electrical Engineer', 'carlos.mendez@soltech-renew.mx', '+52-55-5555-0178', 'Mexico City', 'Mexico'],
        ['Anna', 'Bergstrom', 'Nordic Design Studios', 'UX Designer', 'anna.bergstrom@nordic-ds.se', '+46-8-555-0192', 'Stockholm', 'Sweden'],
        ['James', 'Fletcher', 'Redwood Consulting', 'Senior Consultant', 'james.fletcher@redwoodconsulting.co.uk', '+44-20-5555-0147', 'London', 'United Kingdom'],
        ['Priya', 'Sharma', 'TechVista Solutions', 'Product Manager', 'priya.sharma@techvista.in', '+91-22-5555-0289', 'Mumbai', 'India'],
        ['Oliver', 'Larsen', 'Scandic Maritime AS', 'Fleet Operations Manager', 'oliver.larsen@scandic-maritime.no', '+47-22-555-0134', 'Oslo', 'Norway'],
        ['Maria', 'Santos', 'Verde Agricultural Co', 'Regional Sales Director', 'maria.santos@verde-agri.com.br', '+55-11-5555-0256', 'Sao Paulo', 'Brazil'],
        ['Ahmed', 'Ibrahim', 'Nile Pharma Industries', 'Research Chemist', 'ahmed.ibrahim@nilepharmaceuticals.eg', '+20-2-555-0167', 'Cairo', 'Egypt'],
        ['Sophie', 'Dubois', 'Lumiere Cosmetics', 'Brand Strategist', 'sophie.dubois@lumiere-cosmetiques.fr', '+33-1-5555-0218', 'Paris', 'France'],
        ['Kenji', 'Nakamura', 'Horizon Robotics Inc', 'Mechatronics Engineer', 'kenji.nakamura@horizon-robotics.jp', '+81-45-555-0183', 'Yokohama', 'Japan'],
        ['Grace', 'Mwangi', 'Savanna Tech Hub', 'Community Manager', 'grace.mwangi@savannatech.ke', '+254-20-555-0149', 'Nairobi', 'Kenya'],
        ['Lucas', 'Van den Berg', 'EuroLogistics BV', 'Logistics Coordinator', 'lucas.vandenberg@eurologistics.nl', '+31-20-555-0176', 'Amsterdam', 'Netherlands'],
        ['Natasha', 'Volkov', 'Aurora Digital Agency', 'SEO Specialist', 'natasha.volkov@aurora-digital.ru', '+7-495-555-0231', 'Moscow', 'Russia'],
        ['Benjamin', 'Taylor', 'Summit Ventures Capital', 'Investment Analyst', 'benjamin.taylor@summitvc.com.au', '+61-2-5555-0194', 'Sydney', 'Australia'],
        ['Mei', 'Lin', 'Pacific Rim Trading Co', 'International Buyer', 'mei.lin@pacificrimtrading.tw', '+886-2-5555-0268', 'Taipei', 'Taiwan'],
        ['Daniel', 'Hoffman', 'Apex Cybersecurity GmbH', 'Penetration Tester', 'daniel.hoffman@apex-cybersec.de', '+49-30-555-0152', 'Berlin', 'Germany'],
        ['Chloe', 'Martin', 'Frostbite Game Studios', 'Level Designer', 'chloe.martin@frostbite-gs.ca', '+1-604-555-0287', 'Vancouver', 'Canada'],
        ['Hassan', 'Youssef', 'Crescent Health Services', 'Clinical Data Analyst', 'hassan.youssef@crescent-health.sa', '+966-11-555-0143', 'Riyadh', 'Saudi Arabia'],
        ['Eva', 'Novak', 'Central European Media', 'Broadcast Journalist', 'eva.novak@centraleuromedia.cz', '+420-2-555-0169', 'Prague', 'Czech Republic'],
        ['William', 'Park', 'Hanbit Semiconductor', 'Process Engineer', 'william.park@hanbitsemi.kr', '+82-2-5555-0215', 'Seoul', 'South Korea'],
        ['Aisha', 'Rahman', 'BrightPath Education', 'Curriculum Developer', 'aisha.rahman@brightpath-edu.bd', '+880-2-555-0198', 'Dhaka', 'Bangladesh'],
        ['Michael', 'Andersen', 'Windforce Energy AS', 'Turbine Technician', 'michael.andersen@windforce.dk', '+45-33-555-0174', 'Copenhagen', 'Denmark'],
        ['Lucia', 'Fernandez', 'Andes Mining Corp', 'Geological Surveyor', 'lucia.fernandez@andesmining.cl', '+56-2-555-0241', 'Santiago', 'Chile'],
        ['Patrick', 'OMalley', 'Emerald Isle Breweries', 'Head Brewmaster', 'patrick.omalley@emeraldbreweries.ie', '+353-21-555-0186', 'Cork', 'Ireland'],
        ['Hana', 'Suzuki', 'Zen Garden Wellness', 'Wellness Program Director', 'hana.suzuki@zengardenwellness.jp', '+81-75-555-0253', 'Kyoto', 'Japan'],
        ['Robert', 'Williams', 'Stellar Aerospace Inc', 'Structural Analyst', 'robert.williams@stellaraerospace.com', '+1-281-555-0129', 'Houston', 'United States'],
        ['Ingrid', 'Johansson', 'Arctic Expeditions AB', 'Tour Operations Lead', 'ingrid.johansson@arctic-expeditions.se', '+46-90-555-0167', 'Umea', 'Sweden'],
        ['Francisco', 'Reyes', 'Tropicana Foods SA', 'Quality Control Supervisor', 'francisco.reyes@tropicanafoods.co', '+57-1-555-0238', 'Bogota', 'Colombia'],
        ['Zara', 'Khan', 'Indus Valley Software', 'Backend Developer', 'zara.khan@indusvalley-sw.pk', '+92-21-555-0154', 'Karachi', 'Pakistan'],
        ['Peter', 'Schneider', 'Alpine Instruments AG', 'Optical Engineer', 'peter.schneider@alpine-instruments.ch', '+41-44-555-0189', 'Zurich', 'Switzerland'],
        ['Nadia', 'Popov', 'Eastern Promise Logistics', 'Customs Broker', 'nadia.popov@easternpromise.bg', '+359-2-555-0271', 'Sofia', 'Bulgaria'],
        ['Samuel', 'Adeyemi', 'Baobab Financial Trust', 'Risk Assessment Officer', 'samuel.adeyemi@baobab-financial.gh', '+233-30-555-0146', 'Accra', 'Ghana'],
        ['Christine', 'Lefebvre', 'Chateau Wines Export', 'Export Manager', 'christine.lefebvre@chateauwines.fr', '+33-5-5555-0203', 'Bordeaux', 'France'],
        ['Viktor', 'Horvath', 'Danube Engineering Kft', 'Civil Engineer', 'viktor.horvath@danube-engineering.hu', '+36-1-555-0182', 'Budapest', 'Hungary'],
    ]

    for r, row_data in enumerate(contacts, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set DELIBERATELY WRONG column widths
    # Some too narrow (text will be truncated), some too wide
    ws.column_dimensions['A'].width = 6    # too narrow for "First Name" + data
    ws.column_dimensions['B'].width = 6    # too narrow for "Last Name" + data
    ws.column_dimensions['C'].width = 40   # too wide for Company
    ws.column_dimensions['D'].width = 8    # too narrow for Job Title
    ws.column_dimensions['E'].width = 5    # very narrow for Email (task calls this out)
    ws.column_dimensions['F'].width = 35   # too wide for Phone
    ws.column_dimensions['G'].width = 6    # too narrow for City
    ws.column_dimensions['H'].width = 35   # too wide for Country

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
