"""
Reward Script: Resize all columns A-H to optimal width
Task ID: calc_gfl_057
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4): Previously too-narrow columns (A,B,D,E,G) are now wide enough for content
  Component 2 (0.3): Previously too-wide columns (C,F,H) are now narrower / closer to content width
  Component 3 (0.3): All column widths are proportional to their max content length (optimal fit)
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_057'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(1.0)
            print("PERSIST: ctrl+s sent for libreoffice_calc")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def get_max_content_lengths(ws, max_row, cols):
    """Compute max string length per column across all rows."""
    lengths = {}
    for col_letter in cols:
        col_idx = ord(col_letter) - ord('A') + 1
        max_len = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=col_idx).value
            if v is not None:
                max_len = max(max_len, len(str(v)))
        lengths[col_letter] = max_len
    return lengths


def verify_task(file_path):
    """
    Verify that columns A-H have been resized to optimal width.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Contacts' sheet must exist
    if 'Contacts' not in wb.sheetnames:
        print("FAIL: 'Contacts' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contacts']

    # Precondition: data integrity - 45 rows, 8 columns
    if ws.max_row < 45 or ws.max_column < 8:
        print(f"FAIL: Data integrity - expected >=45 rows and >=8 cols, got {ws.max_row} rows, {ws.max_column} cols")
        print("REWARD: 0.0")
        return 0.0

    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']

    # Known initial widths (before the task)
    initial_widths = {'A': 6.0, 'B': 6.0, 'C': 40.0, 'D': 8.0,
                      'E': 5.0, 'F': 35.0, 'G': 6.0, 'H': 35.0}

    # Get current widths
    current_widths = {}
    for col_letter in cols:
        dim = ws.column_dimensions.get(col_letter)
        if dim and dim.width is not None:
            current_widths[col_letter] = dim.width
        else:
            current_widths[col_letter] = 8.43  # default Excel column width
    print(f"Current widths: {current_widths}")

    # Get max content lengths for reference
    max_lengths = get_max_content_lengths(ws, ws.max_row, cols)
    print(f"Max content lengths: {max_lengths}")

    # Component 1 (0.4 pts): Previously too-narrow columns are now wide enough
    # Columns A(6->10chars), B(6->12chars), D(8->26chars), E(5->38chars), G(6->13chars)
    # were too narrow for their content. After optimal width, they must be wider than content.
    try:
        narrow_cols = ['A', 'B', 'D', 'E', 'G']  # initially too narrow
        widened_count = 0
        for col in narrow_cols:
            # Width must be >= max_content_length (approx 1 char ~ 1 width unit)
            # and must have changed from initial
            min_acceptable = max_lengths[col] - 2  # small tolerance
            if current_widths[col] >= min_acceptable and abs(current_widths[col] - initial_widths[col]) > 0.5:
                print(f"  PASS: Col {col} widened from {initial_widths[col]} to {current_widths[col]} (content needs ~{max_lengths[col]})")
                widened_count += 1
            else:
                print(f"  FAIL: Col {col} width={current_widths[col]}, initial={initial_widths[col]}, content needs ~{max_lengths[col]}")

        if widened_count == len(narrow_cols):
            print(f"PASS: Component 1 - All {len(narrow_cols)} narrow columns widened (0.4 pts)")
            total_score += 0.4
        elif widened_count >= 3:
            partial = 0.4 * (widened_count / len(narrow_cols))
            print(f"PARTIAL: Component 1 - {widened_count}/{len(narrow_cols)} narrow columns widened ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {widened_count}/{len(narrow_cols)} narrow columns widened")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2 (0.3 pts): Previously too-wide columns are now narrower (closer to content)
    # Columns C(40->25chars), F(35->16chars), H(35->20chars) were too wide.
    # After optimal width, they should be reduced and closer to content length.
    try:
        wide_cols = ['C', 'F', 'H']
        narrowed_count = 0
        for col in wide_cols:
            # Must be narrower than initial AND reasonably close to content length
            if current_widths[col] < initial_widths[col] - 1.0 and abs(current_widths[col] - initial_widths[col]) > 1.0:
                print(f"  PASS: Col {col} narrowed from {initial_widths[col]} to {current_widths[col]} (content needs ~{max_lengths[col]})")
                narrowed_count += 1
            else:
                print(f"  FAIL: Col {col} width={current_widths[col]}, initial={initial_widths[col]}, content needs ~{max_lengths[col]}")

        if narrowed_count == len(wide_cols):
            print(f"PASS: Component 2 - All {len(wide_cols)} wide columns narrowed (0.3 pts)")
            total_score += 0.3
        elif narrowed_count >= 2:
            partial = 0.3 * (narrowed_count / len(wide_cols))
            print(f"PARTIAL: Component 2 - {narrowed_count}/{len(wide_cols)} wide columns narrowed ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 - Only {narrowed_count}/{len(wide_cols)} wide columns narrowed")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3 (0.3 pts): All columns have width proportional to content length (optimal fit)
    # Each column width should be within a reasonable range of its max content length.
    # Optimal width typically adds 1-4 chars of padding.
    try:
        optimal_count = 0
        for col in cols:
            content_len = max_lengths[col]
            width = current_widths[col]
            # Acceptable range: content_len - 2 to content_len + 8
            # (LibreOffice optimal width adds some padding)
            if content_len - 2 <= width <= content_len + 8:
                print(f"  PASS: Col {col} width={width} is optimal for content_len={content_len}")
                optimal_count += 1
            else:
                print(f"  FAIL: Col {col} width={width} outside optimal range [{content_len-2}, {content_len+8}] for content_len={content_len}")

        if optimal_count == len(cols):
            print(f"PASS: Component 3 - All {len(cols)} columns have optimal width (0.3 pts)")
            total_score += 0.3
        elif optimal_count >= 6:
            partial = 0.3 * (optimal_count / len(cols))
            print(f"PARTIAL: Component 3 - {optimal_count}/{len(cols)} columns optimal ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 - Only {optimal_count}/{len(cols)} columns have optimal width")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
