"""
Reward Script: Meeting Expense Report with category totals, per-person average,
               named range, currency formatting, and pie chart.
Task ID: calc_wf_021
Domain: libreoffice_calc
Scoring:
  Component 1 — SUMIF formulas for category totals (G2:G5)      0.25
  Component 2 — Per-person average formula in G7                  0.15
  Component 3 — Named range 'ExpenseData' for A1:D31             0.15
  Component 4 — Currency format ($#,##0.00) on amount cells      0.20
  Component 5 — Pie chart with title and percentage labels        0.25
  Total: 1.0
"""

import os
import openpyxl
from openpyxl.chart import PieChart

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_021'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Expenses' not in wb.sheetnames:
        print("CRITICAL: 'Expenses' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Expenses']

    # ---------------------------------------------------------------
    # Component 1: SUMIF formulas for category totals in G2:G5 (0.25)
    # Each of the 4 categories (Meals, Transport, Lodging, Supplies)
    # should have a SUMIF formula in column G.
    # ---------------------------------------------------------------
    try:
        expected_categories = {
            2: 'Meals',
            3: 'Transport',
            4: 'Lodging',
            5: 'Supplies',
        }
        sumif_count = 0
        for row, cat in expected_categories.items():
            val = ws.cell(row=row, column=7).value
            if val and isinstance(val, str) and 'SUMIF' in val.upper():
                # Check it references the right category
                if cat.upper() in val.upper() or cat in val:
                    sumif_count += 1
                    print(f"PASS: G{row} has SUMIF for '{cat}': {val}")
                else:
                    print(f"PARTIAL: G{row} has SUMIF but may not target '{cat}': {val}")
            else:
                print(f"FAIL: G{row} expected SUMIF formula for '{cat}', found: {val}")

        if sumif_count == 4:
            print(f"PASS: Component 1 — All 4 SUMIF formulas present (0.25 pts)")
            total_score += 0.25
        elif sumif_count >= 2:
            partial = round(0.25 * sumif_count / 4, 2)
            print(f"PARTIAL: Component 1 — {sumif_count}/4 SUMIF formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Only {sumif_count}/4 SUMIF formulas found")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Per-person average formula in G7 (0.15)
    # Should compute total / number of unique attendees.
    # ---------------------------------------------------------------
    try:
        g7_val = ws.cell(row=7, column=7).value
        if g7_val and isinstance(g7_val, str) and '=' in g7_val:
            formula_upper = g7_val.upper().replace(' ', '')
            # Should reference G2:G5 sum and count unique attendees
            has_sum_ref = ('SUM' in formula_upper or 'G2' in formula_upper)
            has_count = ('COUNTIF' in formula_upper or 'COUNTA' in formula_upper
                         or 'COUNT' in formula_upper or 'UNIQUE' in formula_upper)
            if has_sum_ref and has_count:
                print(f"PASS: Component 2 — Per-person average formula found: {g7_val} (0.15 pts)")
                total_score += 0.15
            elif has_sum_ref or has_count:
                print(f"PARTIAL: Component 2 — Formula has partial logic: {g7_val} (0.07 pts)")
                total_score += 0.07
            else:
                print(f"FAIL: Component 2 — G7 formula doesn't compute per-person average: {g7_val}")
        else:
            print(f"FAIL: Component 2 — G7 expected a formula, found: {g7_val}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Named range 'ExpenseData' covering A1:D31 (0.15)
    # ---------------------------------------------------------------
    try:
        defined_names = list(wb.defined_names.keys()) if hasattr(wb.defined_names, 'keys') else []
        has_expense_data = False
        correct_range = False

        for name in defined_names:
            if name.upper() == 'EXPENSEDATA':
                has_expense_data = True
                dn = wb.defined_names[name]
                ref = dn.attr_text.upper().replace("'", "").replace('"', '')
                print(f"  Named range '{name}' = {dn.attr_text}")
                # Should reference Expenses!$A$1:$D$31 or similar
                if 'A' in ref and 'D' in ref and '1' in ref and '31' in ref:
                    correct_range = True
                break

        if has_expense_data and correct_range:
            print(f"PASS: Component 3 — Named range 'ExpenseData' with correct range (0.15 pts)")
            total_score += 0.15
        elif has_expense_data:
            print(f"PARTIAL: Component 3 — Named range 'ExpenseData' exists but range may be wrong (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 3 — No named range 'ExpenseData' found. Names: {defined_names}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Currency format ($#,##0.00) on amount cells C2:C31 (0.20)
    # Initial state has 'General' format, golden has '$#,##0.00'.
    # ---------------------------------------------------------------
    try:
        currency_count = 0
        total_cells = 30  # C2:C31
        for r in range(2, 32):
            fmt = ws.cell(row=r, column=3).number_format
            if fmt and '$' in fmt:
                currency_count += 1

        ratio = currency_count / total_cells
        if ratio >= 0.9:
            print(f"PASS: Component 4 — {currency_count}/{total_cells} amount cells have currency format (0.20 pts)")
            total_score += 0.20
        elif ratio >= 0.5:
            partial = round(0.20 * ratio, 2)
            print(f"PARTIAL: Component 4 — {currency_count}/{total_cells} amount cells have currency format ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 4 — Only {currency_count}/{total_cells} amount cells have currency format")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Pie chart showing expense breakdown by category (0.25)
    # Must be a PieChart with relevant title and percentage labels.
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        pie_found = False
        has_title = False
        has_percent = False

        for ch in charts:
            if isinstance(ch, PieChart):
                pie_found = True
                # Check title
                if ch.title:
                    try:
                        # Extract title text from rich text or string
                        title_text = ''
                        if hasattr(ch.title, 'tx') and ch.title.tx:
                            if hasattr(ch.title.tx, 'rich') and ch.title.tx.rich:
                                for p in ch.title.tx.rich.p:
                                    for run in p.r:
                                        title_text += run.t
                            elif hasattr(ch.title.tx, 'strRef'):
                                title_text = str(ch.title.tx.strRef)
                        if title_text:
                            has_title = True
                            print(f"  Pie chart title: '{title_text}'")
                        else:
                            # Title object exists but may not have text
                            has_title = True
                            print(f"  Pie chart has title object (could not extract text)")
                    except Exception:
                        has_title = True
                        print(f"  Pie chart has title (text extraction failed)")

                # Check percentage labels
                if hasattr(ch, 'dataLabels') and ch.dataLabels:
                    if ch.dataLabels.showPercent:
                        has_percent = True
                        print(f"  Pie chart has percentage labels")
                break

        if pie_found and has_title and has_percent:
            print(f"PASS: Component 5 — Pie chart with title and percentage labels (0.25 pts)")
            total_score += 0.25
        elif pie_found and (has_title or has_percent):
            print(f"PARTIAL: Component 5 — Pie chart found but missing title or % labels (0.15 pts)")
            total_score += 0.15
        elif pie_found:
            print(f"PARTIAL: Component 5 — Pie chart found but no title or % labels (0.10 pts)")
            total_score += 0.10
        elif len(charts) > 0:
            print(f"PARTIAL: Component 5 — Chart found but not a PieChart (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — No charts found on 'Expenses' sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
