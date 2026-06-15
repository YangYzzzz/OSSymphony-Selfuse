"""
Initial Setup: Meeting expense report with attendee names, meals, and transport costs.
Task ID: calc_wf_021
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_021'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Expenses'

    # --- Headers ---
    headers = ['Attendee', 'Category', 'Amount', 'Date']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FFD9E2F3", end_color="FFD9E2F3", fill_type="solid")
    header_alignment = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # --- 30 expense entries for 8 attendees ---
    attendees = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'David Kim',
        'Elena Rodriguez', 'James O\'Brien', 'Fatima Al-Rashid', 'Tom Nakamura'
    ]
    categories = ['Meals', 'Transport', 'Lodging', 'Supplies']

    # Realistic expense data - 30 rows
    expense_data = [
        ('Sarah Chen', 'Meals', 45.50, '2025-03-10'),
        ('Marcus Johnson', 'Transport', 32.00, '2025-03-10'),
        ('Priya Patel', 'Lodging', 189.00, '2025-03-10'),
        ('David Kim', 'Meals', 38.75, '2025-03-10'),
        ('Elena Rodriguez', 'Supplies', 27.50, '2025-03-11'),
        ('James O\'Brien', 'Transport', 55.00, '2025-03-11'),
        ('Fatima Al-Rashid', 'Meals', 62.30, '2025-03-11'),
        ('Tom Nakamura', 'Lodging', 189.00, '2025-03-11'),
        ('Sarah Chen', 'Transport', 28.00, '2025-03-11'),
        ('Marcus Johnson', 'Meals', 51.20, '2025-03-12'),
        ('Priya Patel', 'Supplies', 15.80, '2025-03-12'),
        ('David Kim', 'Transport', 42.00, '2025-03-12'),
        ('Elena Rodriguez', 'Meals', 67.90, '2025-03-12'),
        ('James O\'Brien', 'Lodging', 189.00, '2025-03-12'),
        ('Fatima Al-Rashid', 'Supplies', 33.25, '2025-03-12'),
        ('Tom Nakamura', 'Meals', 44.60, '2025-03-13'),
        ('Sarah Chen', 'Lodging', 189.00, '2025-03-13'),
        ('Marcus Johnson', 'Supplies', 22.40, '2025-03-13'),
        ('Priya Patel', 'Meals', 56.75, '2025-03-13'),
        ('David Kim', 'Lodging', 189.00, '2025-03-13'),
        ('Elena Rodriguez', 'Transport', 38.50, '2025-03-14'),
        ('James O\'Brien', 'Meals', 49.80, '2025-03-14'),
        ('Fatima Al-Rashid', 'Transport', 61.00, '2025-03-14'),
        ('Tom Nakamura', 'Supplies', 19.95, '2025-03-14'),
        ('Sarah Chen', 'Meals', 73.25, '2025-03-14'),
        ('Marcus Johnson', 'Transport', 45.00, '2025-03-15'),
        ('Priya Patel', 'Transport', 29.50, '2025-03-15'),
        ('David Kim', 'Supplies', 41.60, '2025-03-15'),
        ('Elena Rodriguez', 'Lodging', 189.00, '2025-03-15'),
        ('Fatima Al-Rashid', 'Meals', 58.40, '2025-03-15'),
    ]

    for r, (attendee, category, amount, date) in enumerate(expense_data, 2):
        ws.cell(row=r, column=1, value=attendee)
        ws.cell(row=r, column=2, value=category)
        ws.cell(row=r, column=3, value=amount)
        ws.cell(row=r, column=4, value=date)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 14

    # --- Summary section at F1 (labels only, NO formulas, NO formatting) ---
    ws['F1'] = 'Category'
    ws['G1'] = 'Total'
    ws['F1'].font = header_font
    ws['G1'].font = header_font
    ws['F1'].fill = header_fill
    ws['G1'].fill = header_fill
    ws['F1'].alignment = header_alignment
    ws['G1'].alignment = header_alignment

    ws['F2'] = 'Meals'
    ws['F3'] = 'Transport'
    ws['F4'] = 'Lodging'
    ws['F5'] = 'Supplies'

    # G2:G5 left EMPTY - no formulas (task asks agent to calculate totals)
    # F7/G7 left EMPTY - no per-person average (task asks agent to add this)
    # NO named range defined (task asks agent to create it)
    # NO currency formatting on amounts (task asks agent to apply it)
    # NO chart (task asks agent to create it)

    ws['F7'] = 'Per-Person Average'
    ws['F7'].font = Font(bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
