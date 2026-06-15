"""
Reward Script: Sales Territory Mapping Sheet
Task ID: calc_gpm_091
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Formulas in E, F, G, H columns for all 8 territories
  Component 2 (0.15): Row 13 National Total SUM formulas + Row 14 National Attainment
  Component 3 (0.10): Title merge A1:H1 with bold 14pt formatting
  Component 4 (0.10): Number formats ($#,##0 for C/D/E/H, % for F/G)
  Component 5 (0.15): Conditional formatting rules on F4:F11, E4:E11, data bars D4:D11
  Component 6 (0.15): Chart exists (col type, 2 series, correct title)
  Component 7 (0.10): Row 13 double bottom border
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_091'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Territory' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Territory' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Territory']

    # =========================================================================
    # Component 1: Formulas in E, F, G, H for rows 4-11 (0.25 points)
    # Initial state has E-H all None. Golden has formulas.
    # =========================================================================
    try:
        formula_count = 0
        total_formula_cells = 0
        for row in range(4, 12):
            # E column: Gap = D - C
            e_val = ws.cell(row=row, column=5).value
            if e_val is not None and isinstance(e_val, str) and '=' in e_val:
                formula_count += 1
            total_formula_cells += 1

            # F column: Attainment = D/C
            f_val = ws.cell(row=row, column=6).value
            if f_val is not None and isinstance(f_val, str) and '=' in f_val:
                formula_count += 1
            total_formula_cells += 1

            # G column: Commission Rate (nested IF)
            g_val = ws.cell(row=row, column=7).value
            if g_val is not None and isinstance(g_val, str) and 'IF' in str(g_val).upper():
                formula_count += 1
            total_formula_cells += 1

            # H column: Commission = D*G
            h_val = ws.cell(row=row, column=8).value
            if h_val is not None and isinstance(h_val, str) and '=' in h_val:
                formula_count += 1
            total_formula_cells += 1

        # Need all 32 formula cells (8 rows x 4 cols)
        ratio = formula_count / total_formula_cells if total_formula_cells > 0 else 0
        if ratio >= 0.9:
            print(f"PASS: Component 1 - Formulas in E/F/G/H ({formula_count}/{total_formula_cells}) (0.25 pts)")
            total_score += 0.25
        elif ratio >= 0.5:
            partial = 0.25 * ratio
            print(f"PARTIAL: Component 1 - Formulas in E/F/G/H ({formula_count}/{total_formula_cells}) ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 - Only {formula_count}/{total_formula_cells} formula cells found")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # =========================================================================
    # Component 2: Row 13 National Total + Row 14 National Attainment (0.15 pts)
    # Initial state: Row 13-14 all None.
    # =========================================================================
    try:
        comp2_score = 0.0

        # Row 13 label
        r13_a = ws.cell(row=13, column=1).value
        if r13_a is not None and 'national' in str(r13_a).lower() and 'total' in str(r13_a).lower():
            comp2_score += 0.02

        # Row 13 SUM formulas in C, D, E, H
        sum_cols = [3, 4, 5, 8]
        sum_found = 0
        for col in sum_cols:
            val = ws.cell(row=13, column=col).value
            if val is not None and isinstance(val, str) and 'SUM' in str(val).upper():
                sum_found += 1
        if sum_found >= 3:
            comp2_score += 0.06
            print(f"  Row 13 SUM formulas: {sum_found}/4 found")
        elif sum_found >= 1:
            comp2_score += 0.03

        # Row 13 bold
        if ws.cell(row=13, column=1).font.bold:
            comp2_score += 0.02

        # Row 14 National Attainment
        r14_a = ws.cell(row=14, column=1).value
        if r14_a is not None and 'attainment' in str(r14_a).lower():
            comp2_score += 0.02

        # Row 14 F column formula (=D13/C13 or similar ratio)
        r14_f = ws.cell(row=14, column=6).value
        if r14_f is not None and isinstance(r14_f, str) and '=' in r14_f:
            comp2_score += 0.03

        if comp2_score >= 0.12:
            print(f"PASS: Component 2 - Row 13 totals + Row 14 attainment ({comp2_score:.3f} pts)")
        else:
            print(f"PARTIAL: Component 2 - Row 13 totals + Row 14 attainment ({comp2_score:.3f} pts)")
        total_score += comp2_score
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # =========================================================================
    # Component 3: Title merge A1:H1 with bold 14pt (0.10 pts)
    # Initial: A1 has text but NOT bold, size 11, NOT merged.
    # Golden: merged A1:H1, bold, size 14.
    # =========================================================================
    try:
        comp3_score = 0.0

        # Check merge
        merged_ranges = [str(r) for r in ws.merged_cells.ranges]
        has_merge = any('A1' in r and 'H1' in r for r in merged_ranges)
        if has_merge:
            comp3_score += 0.04

        # Check bold + size 14
        a1_font = ws['A1'].font
        if a1_font.bold and a1_font.size is not None and a1_font.size >= 14:
            comp3_score += 0.03

        # Check fill color (dark blue)
        try:
            fill_rgb = ws['A1'].fill.fgColor.rgb
            if fill_rgb is not None and '003366' in str(fill_rgb):
                comp3_score += 0.03
        except:
            pass

        if comp3_score >= 0.08:
            print(f"PASS: Component 3 - Title merge + formatting ({comp3_score:.2f} pts)")
        else:
            print(f"PARTIAL: Component 3 - Title merge + formatting ({comp3_score:.2f} pts)")
        total_score += comp3_score
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # =========================================================================
    # Component 4: Number formats (0.10 pts)
    # Initial: all General. Golden: C/D/E/H=$#,##0, F/G=0% or similar %.
    # =========================================================================
    try:
        comp4_score = 0.0

        # Check dollar format on C, D, E, H
        dollar_count = 0
        for col in [3, 4, 5, 8]:
            nf = ws.cell(row=4, column=col).number_format
            if nf is not None and ('$' in str(nf) or '#,##0' in str(nf)):
                dollar_count += 1
        if dollar_count >= 3:
            comp4_score += 0.05

        # Check percent format on F, G
        pct_count = 0
        for col in [6, 7]:
            nf = ws.cell(row=4, column=col).number_format
            if nf is not None and '%' in str(nf):
                pct_count += 1
        if pct_count >= 1:
            comp4_score += 0.05

        if comp4_score >= 0.08:
            print(f"PASS: Component 4 - Number formats ({comp4_score:.2f} pts)")
        else:
            print(f"PARTIAL: Component 4 - Number formats ({comp4_score:.2f} pts)")
        total_score += comp4_score
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # =========================================================================
    # Component 5: Conditional formatting (0.15 pts)
    # Initial: 0 CF rules. Golden: CF on F4:F11, E4:E11, data bars on D4:D11.
    # =========================================================================
    try:
        comp5_score = 0.0

        cf_ranges = {}
        databar_found = False
        for cf in ws.conditional_formatting:
            range_str = str(cf)
            for rule in cf.rules:
                if rule.type == 'dataBar':
                    databar_found = True
                if 'F4' in range_str and 'F11' in range_str:
                    cf_ranges['F'] = cf_ranges.get('F', 0) + 1
                if 'E4' in range_str and 'E11' in range_str:
                    cf_ranges['E'] = cf_ranges.get('E', 0) + 1
                if 'D4' in range_str and 'D11' in range_str:
                    cf_ranges['D'] = cf_ranges.get('D', 0) + 1

        # Attainment CF on F column (should have multiple rules)
        if cf_ranges.get('F', 0) >= 2:
            comp5_score += 0.06
            print(f"  F column CF: {cf_ranges['F']} rules found")

        # Gap CF on E column
        if cf_ranges.get('E', 0) >= 1:
            comp5_score += 0.04
            print(f"  E column CF: {cf_ranges['E']} rules found")

        # Data bars on D column
        if databar_found:
            comp5_score += 0.05
            print(f"  Data bars found on D column")

        if comp5_score >= 0.12:
            print(f"PASS: Component 5 - Conditional formatting ({comp5_score:.2f} pts)")
        else:
            print(f"PARTIAL: Component 5 - Conditional formatting ({comp5_score:.2f} pts)")
        total_score += comp5_score
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # =========================================================================
    # Component 6: Chart (0.15 pts)
    # Initial: 0 charts. Golden: 1 col chart with 2 series.
    # =========================================================================
    try:
        comp6_score = 0.0
        charts = ws._charts

        if len(charts) >= 1:
            comp6_score += 0.05
            ch = charts[0]

            # Check chart type is column/bar
            if ch.type in ('col', 'bar'):
                comp6_score += 0.03

            # Check 2 series (Target vs Actual)
            if len(ch.series) >= 2:
                comp6_score += 0.03

            # Check title contains expected text
            try:
                title_text = ''
                for p in ch.title.tx.rich.p:
                    for r in p.r:
                        title_text += r.t
                if 'territory' in title_text.lower() and 'target' in title_text.lower():
                    comp6_score += 0.04
                    print(f"  Chart title: {title_text}")
            except:
                pass

        if comp6_score >= 0.12:
            print(f"PASS: Component 6 - Chart ({comp6_score:.2f} pts)")
        else:
            print(f"PARTIAL: Component 6 - Chart ({comp6_score:.2f} pts)")
        total_score += comp6_score
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # =========================================================================
    # Component 7: Row 13 double bottom border (0.10 pts)
    # Initial: no special border. Golden: double bottom border on row 13.
    # =========================================================================
    try:
        comp7_score = 0.0
        double_count = 0
        for col in range(1, 9):
            cell = ws.cell(row=13, column=col)
            if cell.border and cell.border.bottom and cell.border.bottom.style == 'double':
                double_count += 1

        if double_count >= 6:
            comp7_score = 0.10
            print(f"PASS: Component 7 - Row 13 double border ({double_count}/8 cols) (0.10 pts)")
        elif double_count >= 3:
            comp7_score = 0.05
            print(f"PARTIAL: Component 7 - Row 13 double border ({double_count}/8 cols) (0.05 pts)")
        else:
            print(f"FAIL: Component 7 - Row 13 double border ({double_count}/8 cols)")
        total_score += comp7_score
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score:.3f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
