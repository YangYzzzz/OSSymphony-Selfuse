"""
Initial Setup: Sales Territory Performance sheet with raw data only.
Task ID: calc_gpm_091
Domain: libreoffice_calc

The initial file contains the Territory sheet with raw data in columns A-D
(Territory, Rep, Target, Actual) but NO formulas, NO formatting, NO charts,
NO totals, NO conditional formatting, and NO data bars.
The agent must create all of those.
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_091'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Territory'

    # Row 1: Plain title text (no merge, no formatting - agent must do that)
    ws['A1'] = 'Sales Territory Performance - Q1 2026'

    # Row 3: Plain headers (no formatting - agent must add bold, colors, borders)
    headers = ['Territory', 'Rep', 'Target', 'Actual', 'Gap', 'Attainment',
               'Commission Rate', 'Commission']
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Rows 4-11: Raw territory data (only columns A-D filled)
    # Columns E-H left empty - agent must add formulas
    territories = [
        ['Northeast',     'Sarah Chen',       520000, 587000],
        ['Southeast',     'Marcus Johnson',    480000, 412000],
        ['Midwest',       'Emily Rodriguez',   390000, 445000],
        ['Southwest',     'David Kim',         350000, 298000],
        ['Northwest',     'Rachel Thompson',   420000, 504000],
        ['Mid-Atlantic',  'James Williams',    600000, 632000],
        ['South Central', 'Lisa Patel',        310000, 275000],
        ['Pacific',       'Michael Chang',     550000, 610000],
    ]

    for r, row_data in enumerate(territories, 4):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths so the data is visible
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Launch LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
