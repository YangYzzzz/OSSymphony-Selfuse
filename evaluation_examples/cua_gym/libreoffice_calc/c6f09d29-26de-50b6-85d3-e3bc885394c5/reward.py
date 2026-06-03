"""
Reward Script: Academic Research Project Tracker in LibreOffice Calc
Task ID: calc_grs_082
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): HYPERLINK formulas in DOI/URL column (F) of Literature Review
  Component 2 (0.20): Conditional formatting on Relevance Score in Literature Review
  Component 3 (0.25): Theme Analysis - COUNTIF formulas for theme source counts (col B)
  Component 4 (0.15): Theme Analysis - Timeline data (D/E cols) and citation statistics (G/H cols)
  Component 5 (0.15): Chart exists on Theme Analysis sheet
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_082'


def persist_app_state(domain: str):
    """Try to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Literature Review sheet must exist
    if 'Literature Review' not in wb.sheetnames:
        print("CRITICAL: 'Literature Review' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws_lr = wb['Literature Review']

    # =========================================================================
    # Component 1: HYPERLINK formulas in DOI/URL column F (0.25 points)
    # In initial_env, column F has plain text URLs. In golden_env, they are
    # wrapped in =HYPERLINK(...) formulas.
    # =========================================================================
    try:
        hyperlink_count = 0
        total_doi_cells = 0
        for row_num in range(2, ws_lr.max_row + 1):
            cell_val = ws_lr.cell(row=row_num, column=6).value  # Column F
            if cell_val is not None:
                total_doi_cells += 1
                if isinstance(cell_val, str) and cell_val.upper().startswith('=HYPERLINK'):
                    hyperlink_count += 1

        if total_doi_cells > 0 and hyperlink_count >= total_doi_cells * 0.8:
            print(f"PASS: Component 1 — {hyperlink_count}/{total_doi_cells} DOI/URL cells have HYPERLINK formulas (0.25 pts)")
            total_score += 0.25
        else:
            print(f"FAIL: Component 1 — Only {hyperlink_count}/{total_doi_cells} DOI/URL cells have HYPERLINK formulas (need >= 80%)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # =========================================================================
    # Component 2: Conditional formatting on Literature Review (0.20 points)
    # Initial has 0 conditional formatting rules. Golden has >= 1 rule applied
    # to the Relevance Score column (H) for green (4-5) and/or gray (1-2).
    # =========================================================================
    try:
        cf_rules = list(ws_lr.conditional_formatting)
        if len(cf_rules) >= 1:
            print(f"PASS: Component 2 — {len(cf_rules)} conditional formatting rule(s) found on Literature Review (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — No conditional formatting rules found on Literature Review")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # =========================================================================
    # Component 3: Theme Analysis - COUNTIF formulas for theme counts (0.25 pts)
    # Initial Theme Analysis has themes listed but NO COUNTIF formulas in col B.
    # Golden has =COUNTIF('Literature Review'!J:J,"*<theme>*") in B2:B9.
    # =========================================================================
    try:
        if 'Theme Analysis' not in wb.sheetnames:
            print("FAIL: Component 3 — 'Theme Analysis' sheet not found")
        else:
            ws_ta = wb['Theme Analysis']
            countif_count = 0
            theme_rows = 0
            for row_num in range(2, ws_ta.max_row + 1):
                theme_val = ws_ta.cell(row=row_num, column=1).value  # Col A
                if theme_val is not None and isinstance(theme_val, str) and len(theme_val.strip()) > 0:
                    theme_rows += 1
                    b_val = ws_ta.cell(row=row_num, column=2).value  # Col B
                    if b_val is not None and isinstance(b_val, str) and 'COUNTIF' in b_val.upper():
                        countif_count += 1

            if theme_rows >= 5 and countif_count >= theme_rows * 0.7:
                print(f"PASS: Component 3 — {countif_count}/{theme_rows} theme rows have COUNTIF formulas (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 3 — {countif_count}/{theme_rows} theme rows have COUNTIF (need >= 70% of >= 5 themes)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # =========================================================================
    # Component 4: Timeline data + Citation statistics on Theme Analysis (0.15 pts)
    # Initial Theme Analysis has no timeline or citation stats columns.
    # Golden has: D/E cols for publication year timeline, G/H cols for citation stats.
    # We check for presence of COUNTIF formulas in E col and/or H col, plus AVERAGE.
    # =========================================================================
    try:
        if 'Theme Analysis' not in wb.sheetnames:
            print("FAIL: Component 4 — 'Theme Analysis' sheet not found")
        else:
            ws_ta = wb['Theme Analysis']
            sub_score = 0.0

            # Check timeline data: D column should have years, E column should have COUNTIF
            timeline_formulas = 0
            for row_num in range(2, 15):
                e_val = ws_ta.cell(row=row_num, column=5).value  # Col E
                if e_val is not None and isinstance(e_val, str) and 'COUNTIF' in e_val.upper():
                    timeline_formulas += 1
            if timeline_formulas >= 3:
                sub_score += 0.075
                print(f"  PASS: Component 4a — {timeline_formulas} timeline COUNTIF formulas in col E")
            else:
                print(f"  FAIL: Component 4a — Only {timeline_formulas} timeline COUNTIF formulas in col E (need >= 3)")

            # Check citation statistics: H column with COUNTIF for source types + AVERAGE
            citation_formulas = 0
            average_formulas = 0
            for row_num in range(2, 20):
                h_val = ws_ta.cell(row=row_num, column=8).value  # Col H
                if h_val is not None and isinstance(h_val, str):
                    if 'COUNTIF' in h_val.upper():
                        citation_formulas += 1
                    if 'AVERAGE' in h_val.upper():
                        average_formulas += 1

            if citation_formulas >= 2 or average_formulas >= 1:
                sub_score += 0.075
                print(f"  PASS: Component 4b — {citation_formulas} citation COUNTIF formulas + {average_formulas} AVERAGE formulas")
            else:
                print(f"  FAIL: Component 4b — {citation_formulas} citation COUNTIF formulas, {average_formulas} AVERAGE formulas")

            if sub_score > 0:
                print(f"PASS: Component 4 — Timeline + citation stats ({sub_score} pts)")
                total_score += sub_score
            else:
                print(f"FAIL: Component 4 — No timeline or citation statistics found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # =========================================================================
    # Component 5: Chart exists on Theme Analysis sheet (0.15 points)
    # Initial has 0 charts. Golden has 1 chart (timeline of sources by year).
    # =========================================================================
    try:
        if 'Theme Analysis' not in wb.sheetnames:
            print("FAIL: Component 5 — 'Theme Analysis' sheet not found")
        else:
            ws_ta = wb['Theme Analysis']
            chart_count = len(ws_ta._charts)
            if chart_count >= 1:
                print(f"PASS: Component 5 — {chart_count} chart(s) found on Theme Analysis (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 5 — No charts found on Theme Analysis sheet")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
