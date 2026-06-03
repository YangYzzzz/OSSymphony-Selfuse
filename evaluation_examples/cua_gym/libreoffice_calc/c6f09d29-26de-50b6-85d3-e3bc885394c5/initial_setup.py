"""
Initial Setup: Academic research project tracker - Literature Review
Task ID: calc_grs_082
Domain: libreoffice_calc

Creates a literature review spreadsheet with realistic ML-in-healthcare papers.
- Literature Review sheet with 18 rows of data, dropdowns for Source Type and Themes
- DOI/URL as plain text (task asks agent to add HYPERLINK formulas)
- NO conditional formatting (task asks agent to apply it)
- Theme Analysis sheet with headers only (task asks agent to add COUNTIF)
- NO chart, NO citation statistics section
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_082'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # ================================================================
    # Sheet 1: Literature Review
    # ================================================================
    ws = wb.active
    ws.title = "Literature Review"

    headers = [
        "Reference ID", "Author(s)", "Title", "Year", "Journal/Publisher",
        "DOI/URL", "Source Type", "Relevance Score", "Key Findings",
        "Themes", "Cited In Your Paper", "Notes"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # Realistic literature review data for ML in healthcare
    data = [
        ["REF-001", "Chen, W.; Patel, R.; Liu, S.", "Deep Learning for Medical Image Segmentation: A Comprehensive Review", 2023, "Nature Reviews Methods Primers", "https://doi.org/10.1038/s43586-023-00218-z", "Journal Article", 5, "Surveys U-Net architectures and transformer-based methods for CT, MRI, and X-ray segmentation. Reports DICE scores above 0.92 for cardiac segmentation.", "Medical Imaging", "Y", "Core reference for Chapter 2"],
        ["REF-002", "Rajpurkar, P.; Chen, E.; Banerjee, O.; Topol, E.J.", "AI in Health and Medicine", 2022, "Nature Medicine", "https://doi.org/10.1038/s41591-021-01614-0", "Journal Article", 5, "Comprehensive overview of clinical AI applications. Highlights regulatory challenges and need for prospective validation studies.", "Clinical Decision Support;Regulatory Issues", "Y", "Foundational review paper"],
        ["REF-003", "Esteva, A.; Kuprel, B.; Novoa, R.A.", "Dermatologist-Level Classification of Skin Cancer with Deep Neural Networks", 2017, "Nature", "https://doi.org/10.1038/nature21056", "Journal Article", 4, "CNN achieves dermatologist-level performance in classifying skin lesions. Trained on 129,450 clinical images.", "Medical Imaging;Deep Learning", "Y", "Seminal paper in medical AI"],
        ["REF-004", "Topol, E.J.", "Deep Medicine: How Artificial Intelligence Can Make Healthcare Human Again", 2019, "Basic Books", "https://doi.org/10.5860/choice.220506", "Book", 3, "Argues AI will free physicians from routine pattern recognition, allowing more time for patient interaction.", "Ethics and Trust;Clinical Decision Support", "N", "Good for introduction framing"],
        ["REF-005", "Beam, A.L.; Kohane, I.S.", "Big Data and Machine Learning in Health Care", 2018, "JAMA", "https://doi.org/10.1001/jama.2017.18391", "Journal Article", 4, "Discusses practical challenges of deploying ML in clinical settings including data quality, interoperability, and validation.", "Data Quality;Clinical Decision Support", "Y", "Important methodology critique"],
        ["REF-006", "Miotto, R.; Wang, F.; Wang, S.; Jiang, X.; Dudley, J.T.", "Deep Learning for Healthcare: Review, Opportunities and Challenges", 2018, "Briefings in Bioinformatics", "https://doi.org/10.1093/bib/bbx044", "Journal Article", 4, "Reviews deep learning applications across EHR, genomics, and imaging. Identifies data heterogeneity as primary barrier.", "Deep Learning;EHR Analysis", "Y", "Strong methods overview"],
        ["REF-007", "Shickel, B.; Tighe, P.J.; Bihorac, A.; Rashidi, P.", "Deep EHR: A Survey of Recent Advances in Deep Learning Techniques for Electronic Health Record Analysis", 2018, "IEEE Journal of Biomedical and Health Informatics", "https://doi.org/10.1109/JBHI.2017.2767063", "Journal Article", 5, "Systematic review of deep learning on EHR data. Covers RNN, CNN, and autoencoder approaches for patient outcome prediction.", "EHR Analysis;Deep Learning", "Y", "Key reference for EHR chapter section"],
        ["REF-008", "Char, D.S.; Shah, N.H.; Magnus, D.", "Implementing Machine Learning in Health Care — Addressing Ethical Challenges", 2018, "New England Journal of Medicine", "https://doi.org/10.1056/NEJMp1714229", "Journal Article", 3, "Discusses bias, transparency, and accountability in clinical ML systems. Proposes framework for ethical evaluation.", "Ethics and Trust;Regulatory Issues", "N", "Ethics discussion reference"],
        ["REF-009", "Liu, X.; Faes, L.; Kale, A.U.; Wagner, S.K.", "A Comparison of Deep Learning Performance Against Health-Care Professionals", 2019, "The Lancet Digital Health", "https://doi.org/10.1016/S2589-7500(19)30123-2", "Journal Article", 4, "Meta-analysis of 82 studies comparing deep learning to clinicians. Finds equivalent diagnostic accuracy but notes study quality concerns.", "Medical Imaging;Deep Learning", "Y", "Strong evidence synthesis"],
        ["REF-010", "Johnson, A.E.W.; Pollard, T.J.; Shen, L.", "MIMIC-III, a Freely Accessible Critical Care Database", 2016, "Scientific Data", "https://doi.org/10.1038/sdata.2016.35", "Journal Article", 3, "Describes the MIMIC-III dataset: 53,423 ICU admissions, de-identified records. Widely used benchmark for clinical ML.", "EHR Analysis;Data Quality", "Y", "Dataset reference"],
        ["REF-011", "Wiens, J.; Saria, S.; Sendak, M.; Ghassemi, M.", "Do No Harm: A Roadmap for Responsible Machine Learning for Health Care", 2019, "Nature Medicine", "https://doi.org/10.1038/s41591-019-0548-6", "Journal Article", 4, "Proposes responsible ML framework: audit trails, ongoing monitoring, and failure mode analysis for clinical deployment.", "Ethics and Trust;Clinical Decision Support", "Y", "Framework for discussion chapter"],
        ["REF-012", "Amann, J.; Blasimme, A.; Vayena, E.; Frey, D.; Madai, V.I.", "Explainability for Artificial Intelligence in Healthcare: A Multidisciplinary Perspective", 2020, "BMC Medical Informatics and Decision Making", "https://doi.org/10.1186/s12911-020-01332-6", "Journal Article", 3, "Reviews XAI methods (LIME, SHAP, attention maps) applied to healthcare. Notes gap between technical and clinical interpretability.", "Ethics and Trust;Deep Learning", "N", "XAI methods overview"],
        ["REF-013", "Ching, T.; Himmelstein, D.S.; Beaulieu-Jones, B.K.", "Opportunities and Obstacles for Deep Learning in Biology and Medicine", 2018, "Journal of The Royal Society Interface", "https://doi.org/10.1098/rsif.2017.0387", "Journal Article", 3, "Broad survey of DL in biology and medicine. Identifies data sharing, reproducibility, and interpretability as key obstacles.", "Deep Learning;Data Quality", "N", "General background"],
        ["REF-014", "Obermeyer, Z.; Powers, B.; Vogeli, C.; Mullainathan, S.", "Dissecting Racial Bias in an Algorithm Used to Manage the Health of Populations", 2019, "Science", "https://doi.org/10.1126/science.aax2342", "Journal Article", 5, "Demonstrates racial bias in widely-used healthcare algorithm affecting 200 million patients. Reducing bias increased Black patient identification by 46.5%.", "Ethics and Trust;Data Quality", "Y", "Critical bias case study"],
        ["REF-015", "Liang, H.; Tsui, B.Y.; Ni, H.; Valentim, C.C.S.", "Evaluation and Accurate Diagnoses of Pediatric Diseases Using Artificial Intelligence", 2019, "Nature Medicine", "https://doi.org/10.1038/s41591-019-0335-4", "Journal Article", 2, "NLP-based system for pediatric diagnosis using Chinese EHR. Limited generalizability to English-language records.", "EHR Analysis;Clinical Decision Support", "N", "Regional study, limited scope"],
        ["REF-016", "McKinney, S.M.; Sieniek, M.; Godbole, V.", "International Evaluation of an AI System for Breast Cancer Screening", 2020, "Nature", "https://doi.org/10.1038/s41586-019-1799-6", "Journal Article", 4, "Google Health AI reduces false positives by 5.7% and false negatives by 9.4% in mammography screening across US and UK datasets.", "Medical Imaging;Clinical Decision Support", "Y", "High-impact screening study"],
        ["REF-017", "Ghassemi, M.; Naumann, T.; Schulam, P.; Beam, A.L.", "Practical Guidance on Artificial Intelligence for Health Care Data", 2019, "The Lancet Digital Health", "https://doi.org/10.1016/S2589-7500(19)30084-6", "Report", 3, "Practical guidelines for working with EHR data: handling missingness, temporal alignment, and label definition.", "Data Quality;EHR Analysis", "N", "Methodology reference"],
        ["REF-018", "Vaswani, A.; Shazeer, N.; Parmar, N.; Uszkoreit, J.", "Attention Is All You Need", 2017, "Advances in Neural Information Processing Systems", "https://doi.org/10.48550/arXiv.1706.03762", "Conference Paper", 2, "Introduces the Transformer architecture. Foundational for later medical NLP and vision transformer applications.", "Deep Learning", "N", "Background architecture reference"],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = header_border
            if c == 8:  # Relevance Score
                cell.alignment = Alignment(horizontal="center")
            if c == 11:  # Cited In Your Paper
                cell.alignment = Alignment(horizontal="center")

    # Column widths
    col_widths = {
        "A": 12, "B": 30, "C": 45, "D": 8, "E": 28,
        "F": 40, "G": 18, "H": 14, "I": 55, "J": 30, "K": 16, "L": 30
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 30

    # Data validation: Source Type dropdown
    dv_source = DataValidation(
        type="list",
        formula1='"Journal Article,Conference Paper,Book,Thesis,Report,Website"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_source.error = "Invalid source type"
    dv_source.errorTitle = "Error"
    dv_source.prompt = "Select source type"
    dv_source.promptTitle = "Source Type"
    dv_source.add("G2:G100")
    ws.add_data_validation(dv_source)

    # Data validation: Themes dropdown
    dv_themes = DataValidation(
        type="list",
        formula1='"Medical Imaging,Deep Learning,EHR Analysis,Clinical Decision Support,Ethics and Trust,Data Quality,Regulatory Issues,NLP in Healthcare"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_themes.error = "Invalid theme"
    dv_themes.errorTitle = "Error"
    dv_themes.prompt = "Select theme"
    dv_themes.promptTitle = "Themes"
    dv_themes.add("J2:J100")
    ws.add_data_validation(dv_themes)

    # Data validation: Cited In Your Paper (Y/N)
    dv_cited = DataValidation(
        type="list",
        formula1='"Y,N"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_cited.add("K2:K100")
    ws.add_data_validation(dv_cited)

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = "A1:L19"

    # ================================================================
    # Sheet 2: Theme Analysis (headers only — task asks agent to add COUNTIF)
    # ================================================================
    ws2 = wb.create_sheet("Theme Analysis")
    ws2["A1"] = "Theme"
    ws2["A1"].font = Font(bold=True)
    ws2["B1"] = "Source Count"
    ws2["B1"].font = Font(bold=True)
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 15

    # List themes as labels (no formulas)
    themes = [
        "Medical Imaging", "Deep Learning", "EHR Analysis",
        "Clinical Decision Support", "Ethics and Trust",
        "Data Quality", "Regulatory Issues", "NLP in Healthcare"
    ]
    for i, theme in enumerate(themes, 2):
        ws2.cell(row=i, column=1, value=theme)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
