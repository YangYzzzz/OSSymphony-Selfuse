"""
Initial Setup: Weather data analysis worksheet - pre-task state
Task ID: calc_wf_056
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from datetime import date, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_056'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Weather'

    # Headers
    headers = ['Date', 'High Temp', 'Low Temp', 'Precipitation', 'Humidity', 'Wind Speed']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 30 days of realistic weather data starting from 2025-06-01
    # Simulating early summer weather with realistic variation
    import random
    random.seed(42)

    start_date = date(2025, 6, 1)
    weather_data = []
    for i in range(30):
        d = start_date + timedelta(days=i)
        # Gradually warming trend with daily variation
        base_high = 28 + (i / 30) * 8  # 28-36 range over the month
        base_low = 16 + (i / 30) * 5   # 16-21 range
        high_temp = round(base_high + random.uniform(-4, 4), 1)
        low_temp = round(base_low + random.uniform(-3, 3), 1)
        # Ensure low < high
        if low_temp >= high_temp:
            low_temp = high_temp - 3.0
        precipitation = round(random.choice([0, 0, 0, 0, 0, 0.2, 0.5, 1.2, 2.8, 5.4, 8.1, 12.3]), 1)
        humidity = random.randint(35, 85)
        wind_speed = round(random.uniform(2, 25), 1)
        weather_data.append([d, high_temp, low_temp, precipitation, humidity, wind_speed])

    for r, row_data in enumerate(weather_data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            if c == 1:
                cell.number_format = 'yyyy-mm-dd'

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
