"""
Reward Script: Weather data analysis worksheet verification
Task ID: calc_wf_056
Domain: libreoffice_calc
Scoring:
  Component 1: Daily Range column (G) with =B-C formulas  — 0.20 pts
  Component 2: Weekly averages summary section             — 0.20 pts
  Component 3: Extremes section (hottest/coldest)          — 0.20 pts
  Component 4: Temperature trend line chart (2 series)     — 0.15 pts
  Component 5: Precipitation bar chart                     — 0.10 pts
  Component 6: Conditional formatting rules                — 0.15 pts
  Total: 1.0
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_wf_056'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Weather' not in wb.sheetnames:
        print("CRITICAL: 'Weather' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Weather']

    # Component 1: Daily Range column (G) with =B-C formulas (0.20 points)
    # Initial file has no column G data; golden adds G1 header + formulas in G2:G31
    try:
        g1_val = ws.cell(row=1, column=7).value
        has_header = g1_val is not None and 'range' in str(g1_val).lower()

        formula_count = 0
        for r in range(2, 32):
            val = ws.cell(row=r, column=7).value
            if val is not None and isinstance(val, str):
                # Check for a formula that computes High - Low (B-C)
                normalized = val.upper().replace(' ', '')
                if normalized.startswith('=') and 'B' in normalized and 'C' in normalized:
                    formula_count += 1

        if has_header and formula_count >= 25:
            print(f"PASS: Component 1 — Daily Range column present with header '{g1_val}' and {formula_count}/30 formulas (0.20 pts)")
            total_score += 0.20
        elif formula_count >= 15:
            print(f"PARTIAL: Component 1 — {formula_count}/30 formulas found but incomplete (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 1 — Expected Daily Range formulas in G2:G31, found {formula_count} formulas, header={g1_val}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Weekly averages summary section (0.20 points)
    # Golden has a summary section with "Summary" label and weekly averages using AVERAGEIFS
    try:
        # Look for "Summary" label and weekly average rows in rows 32-48
        summary_found = False
        averageifs_count = 0

        for r in range(32, 50):
            a_val = ws.cell(row=r, column=1).value
            if a_val is not None and 'summary' in str(a_val).lower():
                summary_found = True

            # Check for AVERAGEIFS formulas in columns B, C, or D
            for c in [2, 3, 4]:
                cell_val = ws.cell(row=r, column=c).value
                if cell_val is not None and isinstance(cell_val, str):
                    if 'AVERAGEIF' in cell_val.upper():
                        averageifs_count += 1

        # Expect summary label + at least some AVERAGEIFS formulas (golden has ~15: 5 weeks x 3 cols)
        if summary_found and averageifs_count >= 8:
            print(f"PASS: Component 2 — Weekly averages section found with 'Summary' label and {averageifs_count} AVERAGEIFS formulas (0.20 pts)")
            total_score += 0.20
        elif averageifs_count >= 3:
            print(f"PARTIAL: Component 2 — Found {averageifs_count} AVERAGEIFS formulas but incomplete (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 2 — Expected weekly averages section, found summary_label={summary_found}, averageifs_count={averageifs_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Extremes section — hottest/coldest day identification (0.20 points)
    # Golden has rows with MAX(B:B), MIN(C:C), INDEX+MATCH for dates
    try:
        extremes_label_found = False
        max_formula_found = False
        min_formula_found = False
        index_match_count = 0

        for r in range(32, 55):
            a_val = ws.cell(row=r, column=1).value
            b_val = ws.cell(row=r, column=2).value

            if a_val is not None and 'extreme' in str(a_val).lower():
                extremes_label_found = True

            if b_val is not None and isinstance(b_val, str):
                upper_val = b_val.upper().replace(' ', '')
                if '=MAX(' in upper_val:
                    max_formula_found = True
                if '=MIN(' in upper_val:
                    min_formula_found = True
                if 'INDEX(' in upper_val and 'MATCH(' in upper_val:
                    index_match_count += 1

        score_3 = 0.0
        if max_formula_found and min_formula_found:
            score_3 += 0.10
        if index_match_count >= 2:
            score_3 += 0.10

        if score_3 > 0:
            total_score += score_3
            print(f"PASS: Component 3 — Extremes section: MAX={max_formula_found}, MIN={min_formula_found}, INDEX/MATCH={index_match_count} ({score_3} pts)")
        else:
            print(f"FAIL: Component 3 — Expected extremes formulas (MAX, MIN, INDEX/MATCH). Found MAX={max_formula_found}, MIN={min_formula_found}, INDEX/MATCH={index_match_count}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Temperature trend line chart with 2 series (High and Low) (0.15 points)
    # Initial file has 0 charts; golden has a LineChart titled "Temperature Trend" with 2 series
    try:
        charts = ws._charts
        line_chart_found = False
        line_series_count = 0

        for chart in charts:
            class_name = chart.__class__.__name__
            if class_name == 'LineChart':
                line_chart_found = True
                line_series_count = len(chart.series)
                break

        if line_chart_found and line_series_count >= 2:
            print(f"PASS: Component 4 — Line chart found with {line_series_count} series (0.15 pts)")
            total_score += 0.15
        elif line_chart_found:
            print(f"PARTIAL: Component 4 — Line chart found but only {line_series_count} series (0.08 pts)")
            total_score += 0.08
        else:
            print(f"FAIL: Component 4 — No LineChart found. Total charts on sheet: {len(charts)}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: Precipitation bar chart (0.10 points)
    # Golden has a BarChart titled "Daily Precipitation" with 1 series referencing column D
    try:
        charts = ws._charts
        bar_chart_found = False

        for chart in charts:
            class_name = chart.__class__.__name__
            if class_name == 'BarChart':
                bar_chart_found = True
                break

        if bar_chart_found:
            print(f"PASS: Component 5 — Bar chart (precipitation) found (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No BarChart found. Total charts on sheet: {len(charts)}")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Conditional formatting rules (0.15 points)
    # Golden has 2 CF rules: B2:B31 > 35 (red fill), C2:C31 < 0 (blue fill)
    # Initial file has 0 CF rules
    try:
        cf_rules = list(ws.conditional_formatting)
        high_temp_cf = False
        low_temp_cf = False

        for cf in cf_rules:
            range_str = str(cf)
            for rule in cf.rules:
                operator = getattr(rule, 'operator', None)
                formula = getattr(rule, 'formula', None)

                # Check for High Temp > 35 rule on column B
                if operator == 'greaterThan' and formula:
                    for f in formula:
                        if '35' in str(f):
                            high_temp_cf = True

                # Check for Low Temp < 0 rule on column C
                if operator == 'lessThan' and formula:
                    for f in formula:
                        if '0' in str(f):
                            low_temp_cf = True

        score_6 = 0.0
        if high_temp_cf:
            score_6 += 0.075
        if low_temp_cf:
            score_6 += 0.075

        if score_6 > 0:
            total_score += score_6
            print(f"PASS: Component 6 — Conditional formatting: high_temp_rule={high_temp_cf}, low_temp_rule={low_temp_cf} ({score_6} pts)")
        else:
            print(f"FAIL: Component 6 — No matching conditional formatting rules found. Total CF ranges: {len(cf_rules)}")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
