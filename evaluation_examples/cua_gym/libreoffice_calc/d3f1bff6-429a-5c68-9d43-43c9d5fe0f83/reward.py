"""
Reward Script: Employee Directory with sorting, HYPERLINK emails, date formatting,
               freeze panes, auto-filter, and COUNTIF department headcount summary.
Task ID: calc_grs_009
Domain: libreoffice_calc
Scoring:
  Component 1 — Data sorted by Department then Last Name (0.25)
  Component 2 — HYPERLINK mailto formulas in Email column (0.20)
  Component 3 — Custom date format 'MMM DD, YYYY' on Start Date column (0.15)
  Component 4 — Freeze panes on header row (0.10)
  Component 5 — AutoFilter on all columns (0.10)
  Component 6 — Department headcount summary with COUNTIF formulas (0.20)
"""

import os
import datetime
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_009'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Find the main sheet (should be the first/only sheet with employee data)
    ws = None
    for sn in wb.sheetnames:
        sheet = wb[sn]
        if sheet.cell(row=1, column=1).value and 'Employee' in str(sheet.cell(row=1, column=1).value):
            ws = sheet
            break
    if ws is None:
        ws = wb.active

    # Determine data range: find last row with employee data (column A starts with EMP-)
    last_data_row = 1
    for r in range(2, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and str(val).startswith('EMP-'):
            last_data_row = r
        else:
            break

    num_employees = last_data_row - 1  # subtract header
    print(f"INFO: Found {num_employees} employees (rows 2-{last_data_row})")

    if num_employees < 20:
        print(f"WARN: Task requires at least 20 employees, found {num_employees}")

    # Component 1: Data sorted by Department then by Last Name (0.25 points)
    # This checks that the data is sorted first by Department (col C) alphabetically,
    # then within each department by last name alphabetically.
    try:
        departments = []
        names_by_dept = {}
        for r in range(2, last_data_row + 1):
            dept = ws.cell(row=r, column=3).value
            full_name = ws.cell(row=r, column=2).value
            if dept and full_name:
                departments.append(dept)
                if dept not in names_by_dept:
                    names_by_dept[dept] = []
                # Extract last name (last word of full name)
                last_name = str(full_name).strip().split()[-1]
                names_by_dept[dept].append(last_name)

        # Check departments are sorted alphabetically
        unique_depts_in_order = []
        for d in departments:
            if not unique_depts_in_order or unique_depts_in_order[-1] != d:
                unique_depts_in_order.append(d)

        depts_sorted = unique_depts_in_order == sorted(unique_depts_in_order)

        # Check within each department, last names are sorted
        names_not_sorted_dept = None
        for dept, last_names in names_by_dept.items():
            if last_names != sorted(last_names, key=str.lower):
                names_not_sorted_dept = dept
                print(f"  DETAIL: Dept '{dept}' not sorted by last name: {last_names}")
                break
        names_sorted = (names_not_sorted_dept is None)

        # Also verify that departments are grouped (not interleaved)
        interleave_found = None
        seen_depts = set()
        prev_dept = None
        for d in departments:
            if d != prev_dept:
                if d in seen_depts:
                    interleave_found = d
                    break
                seen_depts.add(d)
            prev_dept = d
        dept_groups_clean = (interleave_found is None)

        if depts_sorted and names_sorted and dept_groups_clean:
            print(f"PASS: Component 1 — Data sorted by Department then Last Name (0.25 pts)")
            total_score += 0.25
        elif depts_sorted and dept_groups_clean:
            # Partial: departments sorted but names within not fully sorted
            print(f"PARTIAL: Component 1 — Departments sorted but names within not fully sorted (0.12 pts)")
            total_score += 0.12
        else:
            print(f"FAIL: Component 1 — Data not sorted. Depts sorted: {depts_sorted}, grouped: {dept_groups_clean}, names sorted: {names_sorted}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: HYPERLINK mailto formulas in Email column (0.20 points)
    # Initial has plain text emails; golden has =HYPERLINK("mailto:...","...")
    try:
        hyperlink_count = 0
        total_emails = 0
        for r in range(2, last_data_row + 1):
            cell_val = ws.cell(row=r, column=5).value
            if cell_val is not None:
                total_emails += 1
                val_str = str(cell_val)
                if val_str.upper().startswith('=HYPERLINK') and 'mailto:' in val_str.lower():
                    hyperlink_count += 1

        if total_emails > 0 and hyperlink_count == total_emails:
            print(f"PASS: Component 2 — All {hyperlink_count}/{total_emails} emails have HYPERLINK mailto formulas (0.20 pts)")
            total_score += 0.20
        elif hyperlink_count > 0:
            ratio = hyperlink_count / max(total_emails, 1)
            partial = round(0.20 * ratio, 2)
            if partial > 0:  # award partial credit
                print(f"PARTIAL: Component 2 — {hyperlink_count}/{total_emails} emails have HYPERLINK ({partial} pts)")
                total_score += partial
        else:
            print(f"FAIL: Component 2 — No HYPERLINK mailto formulas found in email column. Sample: {ws.cell(row=2, column=5).value}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Custom date format 'MMM DD, YYYY' on Start Date column (0.15 points)
    # Initial has plain text dates with 'General' format; golden has datetime with 'MMM DD, YYYY'
    try:
        date_format_count = 0
        date_is_datetime_count = 0
        total_dates = 0
        for r in range(2, last_data_row + 1):
            cell = ws.cell(row=r, column=8)
            if cell.value is not None:
                total_dates += 1
                # Check if value is datetime (not string)
                if isinstance(cell.value, datetime.datetime):
                    date_is_datetime_count += 1
                # Check number format contains MMM
                nf = str(cell.number_format).upper()
                if 'MMM' in nf:
                    date_format_count += 1

        if total_dates > 0 and date_format_count == total_dates and date_is_datetime_count == total_dates:
            print(f"PASS: Component 3 — All {date_format_count}/{total_dates} dates have MMM format and are datetime objects (0.15 pts)")
            total_score += 0.15
        elif date_is_datetime_count == total_dates and date_is_datetime_count > 0:
            # Dates are datetime but format may differ
            print(f"PARTIAL: Component 3 — Dates are datetime but format not 'MMM DD, YYYY' ({date_format_count}/{total_dates} match) (0.07 pts)")
            total_score += 0.07
        else:
            sample_cell = ws.cell(row=2, column=8)
            print(f"FAIL: Component 3 — Dates not in custom format. Sample: value={sample_cell.value}, type={type(sample_cell.value).__name__}, format={sample_cell.number_format}")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Freeze panes on header row (0.10 points)
    # Initial has freeze_panes=None; golden has freeze_panes='A2'
    try:
        fp = ws.freeze_panes
        if fp is not None and str(fp) == 'A2':
            print(f"PASS: Component 4 — Freeze panes set to A2 (header row frozen) (0.10 pts)")
            total_score += 0.10
        elif fp is not None:
            # Some freeze is set, partial credit
            print(f"PARTIAL: Component 4 — Freeze panes set to {fp}, expected A2 (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No freeze panes set (expected A2)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: AutoFilter on all columns (0.10 points)
    # Initial has no auto_filter; golden has auto_filter on the data range
    try:
        af = ws.auto_filter.ref
        if af is not None and af != '':
            # Verify the filter covers all 9 columns and at least the header+data rows
            print(f"PASS: Component 5 — AutoFilter set to {af} (0.10 pts)")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — No AutoFilter set")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # Component 6: Department headcount summary with COUNTIF formulas (0.20 points)
    # Initial has no summary section; golden has rows 25-31 with department names and COUNTIF
    try:
        # Look for a summary section below the data
        summary_found = False
        countif_count = 0
        dept_names_in_summary = []

        for r in range(last_data_row + 1, ws.max_row + 1):
            a_val = ws.cell(row=r, column=1).value
            b_val = ws.cell(row=r, column=2).value

            if a_val and 'summary' in str(a_val).lower():
                summary_found = (a_val is not None)  # derived from actual cell check

            if b_val and str(b_val).upper().startswith('=COUNTIF'):
                countif_count += 1
                if a_val:
                    dept_names_in_summary.append(str(a_val))

        # We expect at least some COUNTIF formulas for department headcount
        if summary_found and countif_count >= 3:
            print(f"PASS: Component 6 — Summary section found with {countif_count} COUNTIF formulas for departments: {dept_names_in_summary} (0.20 pts)")
            total_score += 0.20
        elif countif_count >= 1:
            partial = round(0.20 * min(countif_count / 3.0, 1.0), 2)
            print(f"PARTIAL: Component 6 — Found {countif_count} COUNTIF formulas but summary section incomplete ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 6 — No department headcount summary with COUNTIF formulas found below data")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
