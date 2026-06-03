"""
Initial Setup: Employee Directory Spreadsheet
Task ID: calc_grs_009
Domain: libreoffice_calc
Creates an unsorted employee directory with plain data - no sorting, no hyperlinks,
no custom date format, no freeze panes, no autofilter, no summary formulas.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_009'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Directory"

    # Headers
    headers = [
        "Employee ID", "Full Name", "Department", "Job Title",
        "Email", "Phone", "Location", "Start Date", "Manager"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Deliberately UNSORTED employee data (mixed departments, not alphabetical)
    employees = [
        ["EMP-1001", "Sarah Chen", "Engineering", "Senior Software Engineer",
         "sarah.chen@acmecorp.com", "(415) 555-0142", "San Francisco HQ", "2021-03-15", "David Park"],
        ["EMP-1002", "Marcus Johnson", "Marketing", "Digital Marketing Manager",
         "marcus.johnson@acmecorp.com", "(212) 555-0198", "New York", "2022-06-01", "Linda Torres"],
        ["EMP-1003", "Priya Patel", "Finance", "Financial Analyst",
         "priya.patel@acmecorp.com", "(312) 555-0267", "Chicago", "2023-01-10", "Robert Kim"],
        ["EMP-1004", "James Wright", "Engineering", "DevOps Engineer",
         "james.wright@acmecorp.com", "(415) 555-0311", "San Francisco HQ", "2020-08-22", "David Park"],
        ["EMP-1005", "Olivia Martinez", "Human Resources", "HR Coordinator",
         "olivia.martinez@acmecorp.com", "(512) 555-0423", "Austin", "2022-11-05", "Nancy Liu"],
        ["EMP-1006", "Wei Zhang", "Engineering", "Frontend Developer",
         "wei.zhang@acmecorp.com", "(415) 555-0189", "San Francisco HQ", "2023-07-18", "Sarah Chen"],
        ["EMP-1007", "Emily Brooks", "Sales", "Account Executive",
         "emily.brooks@acmecorp.com", "(212) 555-0354", "New York", "2021-09-30", "Carlos Rivera"],
        ["EMP-1008", "Ahmed Hassan", "Finance", "Senior Accountant",
         "ahmed.hassan@acmecorp.com", "(312) 555-0476", "Chicago", "2019-04-12", "Robert Kim"],
        ["EMP-1009", "Jessica Taylor", "Marketing", "Content Strategist",
         "jessica.taylor@acmecorp.com", "(512) 555-0538", "Austin", "2023-03-25", "Marcus Johnson"],
        ["EMP-1010", "David Park", "Engineering", "Engineering Manager",
         "david.park@acmecorp.com", "(415) 555-0621", "San Francisco HQ", "2018-01-08", "VP Engineering"],
        ["EMP-1011", "Sophia Nguyen", "Sales", "Sales Director",
         "sophia.nguyen@acmecorp.com", "(212) 555-0745", "New York", "2019-11-20", "VP Sales"],
        ["EMP-1012", "Carlos Rivera", "Sales", "Regional Sales Manager",
         "carlos.rivera@acmecorp.com", "(312) 555-0867", "Chicago", "2020-05-14", "Sophia Nguyen"],
        ["EMP-1013", "Rachel Green", "Human Resources", "Talent Acquisition Lead",
         "rachel.green@acmecorp.com", "(512) 555-0912", "Austin", "2021-02-28", "Nancy Liu"],
        ["EMP-1014", "Robert Kim", "Finance", "Finance Director",
         "robert.kim@acmecorp.com", "(312) 555-1034", "Chicago", "2017-06-15", "CFO"],
        ["EMP-1015", "Linda Torres", "Marketing", "VP Marketing",
         "linda.torres@acmecorp.com", "(212) 555-1156", "New York", "2018-09-03", "CEO"],
        ["EMP-1016", "Nathan Cooper", "Engineering", "QA Engineer",
         "nathan.cooper@acmecorp.com", "(415) 555-1278", "San Francisco HQ", "2022-04-19", "David Park"],
        ["EMP-1017", "Mia Anderson", "Finance", "Payroll Specialist",
         "mia.anderson@acmecorp.com", "(312) 555-1390", "Chicago", "2023-08-07", "Ahmed Hassan"],
        ["EMP-1018", "Tyler Washington", "Sales", "Business Development Rep",
         "tyler.washington@acmecorp.com", "(512) 555-1412", "Austin", "2024-01-15", "Carlos Rivera"],
        ["EMP-1019", "Nancy Liu", "Human Resources", "HR Director",
         "nancy.liu@acmecorp.com", "(415) 555-1534", "San Francisco HQ", "2019-07-22", "VP Operations"],
        ["EMP-1020", "Alexander Mitchell", "Marketing", "Graphic Designer",
         "alexander.mitchell@acmecorp.com", "(212) 555-1656", "New York", "2022-12-11", "Jessica Taylor"],
        ["EMP-1021", "Isabelle Dupont", "Engineering", "Data Engineer",
         "isabelle.dupont@acmecorp.com", "(415) 555-1789", "San Francisco HQ", "2023-05-02", "David Park"],
        ["EMP-1022", "Derek Sullivan", "Sales", "Account Manager",
         "derek.sullivan@acmecorp.com", "(312) 555-1823", "Chicago", "2021-10-18", "Carlos Rivera"],
    ]

    for r, emp in enumerate(employees, 2):
        for c, val in enumerate(emp, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths for readability
    col_widths = {
        'A': 14, 'B': 22, 'C': 18, 'D': 28,
        'E': 34, 'F': 18, 'G': 22, 'H': 14, 'I': 20
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc for GUI-ready state
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
