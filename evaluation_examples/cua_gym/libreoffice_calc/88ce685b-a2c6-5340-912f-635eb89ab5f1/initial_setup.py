"""
Initial Setup: Insert 3 new sheets at the end of the workbook
Task ID: calc_ps_067
Domain: libreoffice_calc

Creates a workbook with sheets 'July', 'August', 'September',
each containing realistic monthly sales data.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_ps_067'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # Shared styling
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    headers = ['Date', 'Sales Rep', 'Product', 'Units Sold', 'Unit Price', 'Total Revenue']

    # --- July Data ---
    july_data = [
        ['2025-07-02', 'Sarah Chen', 'Widget Pro', 15, 29.99, 449.85],
        ['2025-07-05', 'Marcus Johnson', 'GadgetX', 8, 74.50, 596.00],
        ['2025-07-07', 'Elena Rodriguez', 'Widget Pro', 22, 29.99, 659.78],
        ['2025-07-10', 'David Park', 'SmartModule', 5, 149.00, 745.00],
        ['2025-07-12', 'Sarah Chen', 'GadgetX', 12, 74.50, 894.00],
        ['2025-07-15', 'Aisha Patel', 'Widget Pro', 30, 29.99, 899.70],
        ['2025-07-18', 'Marcus Johnson', 'SmartModule', 3, 149.00, 447.00],
        ['2025-07-20', 'Elena Rodriguez', 'GadgetX', 10, 74.50, 745.00],
        ['2025-07-23', 'David Park', 'Widget Pro', 18, 29.99, 539.82],
        ['2025-07-25', 'Aisha Patel', 'SmartModule', 7, 149.00, 1043.00],
        ['2025-07-28', 'Sarah Chen', 'GadgetX', 14, 74.50, 1043.00],
        ['2025-07-30', 'Marcus Johnson', 'Widget Pro', 25, 29.99, 749.75],
    ]

    # --- August Data ---
    august_data = [
        ['2025-08-01', 'Elena Rodriguez', 'SmartModule', 6, 149.00, 894.00],
        ['2025-08-04', 'David Park', 'GadgetX', 9, 74.50, 670.50],
        ['2025-08-06', 'Sarah Chen', 'Widget Pro', 20, 29.99, 599.80],
        ['2025-08-08', 'Aisha Patel', 'GadgetX', 11, 74.50, 819.50],
        ['2025-08-11', 'Marcus Johnson', 'SmartModule', 4, 149.00, 596.00],
        ['2025-08-14', 'Elena Rodriguez', 'Widget Pro', 28, 29.99, 839.72],
        ['2025-08-16', 'David Park', 'SmartModule', 8, 149.00, 1192.00],
        ['2025-08-19', 'Sarah Chen', 'GadgetX', 16, 74.50, 1192.00],
        ['2025-08-22', 'Aisha Patel', 'Widget Pro', 13, 29.99, 389.87],
        ['2025-08-25', 'Marcus Johnson', 'GadgetX', 7, 74.50, 521.50],
        ['2025-08-27', 'Elena Rodriguez', 'SmartModule', 10, 149.00, 1490.00],
        ['2025-08-30', 'David Park', 'Widget Pro', 19, 29.99, 569.81],
    ]

    # --- September Data ---
    september_data = [
        ['2025-09-02', 'Aisha Patel', 'GadgetX', 13, 74.50, 968.50],
        ['2025-09-04', 'Sarah Chen', 'SmartModule', 5, 149.00, 745.00],
        ['2025-09-07', 'Marcus Johnson', 'Widget Pro', 24, 29.99, 719.76],
        ['2025-09-09', 'Elena Rodriguez', 'GadgetX', 10, 74.50, 745.00],
        ['2025-09-12', 'David Park', 'Widget Pro', 17, 29.99, 509.83],
        ['2025-09-15', 'Aisha Patel', 'SmartModule', 9, 149.00, 1341.00],
        ['2025-09-17', 'Sarah Chen', 'Widget Pro', 21, 29.99, 629.79],
        ['2025-09-19', 'Marcus Johnson', 'GadgetX', 15, 74.50, 1117.50],
        ['2025-09-22', 'Elena Rodriguez', 'SmartModule', 6, 149.00, 894.00],
        ['2025-09-24', 'David Park', 'GadgetX', 8, 74.50, 596.00],
        ['2025-09-26', 'Aisha Patel', 'Widget Pro', 26, 29.99, 779.74],
        ['2025-09-29', 'Sarah Chen', 'SmartModule', 11, 149.00, 1639.00],
    ]

    sheet_configs = [
        ('July', july_data),
        ('August', august_data),
        ('September', september_data),
    ]

    for idx, (sheet_name, data) in enumerate(sheet_configs):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)

        # Write headers
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

        # Write data
        for r, row_data in enumerate(data, 2):
            for c, val in enumerate(row_data, 1):
                cell = ws.cell(row=r, column=c, value=val)
                cell.border = thin_border
                if c in (5, 6):  # Unit Price, Total Revenue
                    cell.number_format = '$#,##0.00'
                elif c == 4:  # Units Sold
                    cell.number_format = '0'

        # Set column widths
        col_widths = [12, 20, 16, 12, 12, 16]
        for i, w in enumerate(col_widths):
            ws.column_dimensions[chr(65 + i)].width = w

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
