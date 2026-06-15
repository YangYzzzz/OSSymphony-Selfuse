"""
Initial Setup: Create a response times spreadsheet with no conditional formatting
Task ID: calc_gcv_019
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gcv_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Response_Times"

    # Headers
    headers = [
        "Ticket ID", "Category", "Priority", "Agent",
        "Created", "Resolved", "Response Time (hours)"
    ]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22

    # Data
    categories = [
        "Billing", "Technical Support", "Account Access", "Feature Request",
        "Bug Report", "Onboarding", "Security", "Performance", "Data Export",
        "Integration"
    ]
    priorities = ["Critical", "High", "Medium", "Low"]
    agents = [
        "Sarah Chen", "Marcus Johnson", "Elena Rodriguez", "David Kim",
        "Priya Patel", "James O'Brien", "Aisha Mohammed", "Carlos Rivera",
        "Yuki Tanaka", "Lisa Thompson"
    ]

    # Generate 49 rows of realistic data
    base_year = 2025
    for i in range(49):
        row = i + 2
        ticket_id = f"TK-{10001 + i}"
        category = random.choice(categories)
        priority = random.choice(priorities)
        agent = random.choice(agents)

        # Created date: spread across Jan-Mar 2025
        month = random.randint(1, 3)
        day = random.randint(1, 28)
        hour = random.randint(8, 18)
        minute = random.choice([0, 15, 30, 45])
        created = f"{base_year}-{month:02d}-{day:02d} {hour:02d}:{minute:02d}"

        # Resolved: 0.5 to 168 hours later (matching context range)
        # Distribution: mostly short, some long
        if priority == "Critical":
            resp_time = round(random.uniform(0.5, 8.0), 1)
        elif priority == "High":
            resp_time = round(random.uniform(1.0, 24.0), 1)
        elif priority == "Medium":
            resp_time = round(random.uniform(4.0, 72.0), 1)
        else:  # Low
            resp_time = round(random.uniform(12.0, 168.0), 1)

        # Compute resolved datetime (approximate for display)
        resolved_hour = hour + int(resp_time)
        resolved_day = day + resolved_hour // 24
        resolved_hour = resolved_hour % 24
        resolved_day = min(resolved_day, 28)
        resolved_month = month
        if resolved_day > 28:
            resolved_day = resolved_day - 28
            resolved_month = min(month + 1, 3)
        resolved = f"{base_year}-{resolved_month:02d}-{resolved_day:02d} {resolved_hour:02d}:{minute:02d}"

        ws.cell(row=row, column=1, value=ticket_id)
        ws.cell(row=row, column=2, value=category)
        ws.cell(row=row, column=3, value=priority)
        ws.cell(row=row, column=4, value=agent)
        ws.cell(row=row, column=5, value=created)
        ws.cell(row=row, column=6, value=resolved)
        cell_g = ws.cell(row=row, column=7, value=resp_time)
        cell_g.number_format = '0.0'

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting on initial file

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
