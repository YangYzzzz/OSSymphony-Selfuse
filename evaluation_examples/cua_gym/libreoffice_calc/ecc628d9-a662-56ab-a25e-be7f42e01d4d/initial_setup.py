"""
Initial Setup: Create StoreSales sheet with 480 transaction rows
Task ID: calc_pivot_034
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_034'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "StoreSales"

    # Headers
    headers = ["TxnID", "TxnDate", "StoreID", "Category", "SalesAmount"]
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 16

    stores = ["Store1", "Store2", "Store3", "Store4"]
    categories = [
        "Electronics", "Clothing", "Groceries", "Home & Garden",
        "Sports", "Books", "Toys", "Automotive", "Health", "Office Supplies"
    ]

    # We need 480 rows: 30 per store per quarter (4 stores x 4 quarters x 30 = 480)
    # Target totals:
    #   Q1/Store1 = 28000
    #   Q2/Store2 = 32000
    #   Grand total = 520000
    #
    # Design: 16 buckets (4 quarters x 4 stores), each 30 txns.
    # Assign target sums per bucket, ensuring they add to 520000.
    # Q1/Store1=28000, Q2/Store2=32000, remaining 14 buckets share 520000-28000-32000=460000.
    # Average per remaining bucket ~ 32857. Let's distribute with some variance.

    bucket_targets = {
        ("Q1", "Store1"): 28000,
        ("Q1", "Store2"): 31500,
        ("Q1", "Store3"): 29800,
        ("Q1", "Store4"): 33200,
        ("Q2", "Store1"): 34500,
        ("Q2", "Store2"): 32000,
        ("Q2", "Store3"): 35100,
        ("Q2", "Store4"): 30900,
        ("Q3", "Store1"): 33800,
        ("Q3", "Store2"): 34200,
        ("Q3", "Store3"): 31600,
        ("Q3", "Store4"): 32400,
        ("Q4", "Store1"): 36000,
        ("Q4", "Store2"): 35500,
        ("Q4", "Store3"): 33700,
        ("Q4", "Store4"): 27800,
    }

    # Verify grand total
    assert sum(bucket_targets.values()) == 520000, f"Total: {sum(bucket_targets.values())}"

    # Quarter date ranges
    quarter_date_ranges = {
        "Q1": (date(2024, 1, 1), date(2024, 3, 31)),
        "Q2": (date(2024, 4, 1), date(2024, 6, 30)),
        "Q3": (date(2024, 7, 1), date(2024, 9, 30)),
        "Q4": (date(2024, 10, 1), date(2024, 12, 31)),
    }

    def random_date_in_range(start, end):
        delta = (end - start).days
        return start + timedelta(days=random.randint(0, delta))

    def generate_amounts(target_sum, count):
        """Generate 'count' amounts that sum exactly to target_sum."""
        # Generate random proportions, then scale
        raw = [random.uniform(0.5, 2.0) for _ in range(count)]
        total_raw = sum(raw)
        # Scale to target, round to 2 decimals
        amounts = [round(target_sum * r / total_raw, 2) for r in raw]
        # Fix rounding error on last element
        diff = round(target_sum - sum(amounts), 2)
        amounts[-1] = round(amounts[-1] + diff, 2)
        return amounts

    # Build all 480 rows
    rows = []
    txn_id = 1
    for quarter in ["Q1", "Q2", "Q3", "Q4"]:
        for store in stores:
            target = bucket_targets[(quarter, store)]
            amounts = generate_amounts(target, 30)
            start_d, end_d = quarter_date_ranges[quarter]
            for i in range(30):
                txn_date = random_date_in_range(start_d, end_d)
                category = random.choice(categories)
                rows.append((txn_id, txn_date, store, category, amounts[i]))
                txn_id += 1

    # Shuffle rows to make it realistic (not grouped by quarter/store)
    random.shuffle(rows)

    # Re-assign TxnIDs sequentially after shuffle
    date_fmt = Font(name="Calibri", size=11)
    money_fmt = '#,##0.00'
    for idx, (_, txn_date, store, category, amount) in enumerate(rows):
        row_num = idx + 2
        ws.cell(row=row_num, column=1, value=idx + 1)
        date_cell = ws.cell(row=row_num, column=2, value=txn_date)
        date_cell.number_format = 'yyyy-mm-dd'
        ws.cell(row=row_num, column=3, value=store)
        ws.cell(row=row_num, column=4, value=category)
        amt_cell = ws.cell(row=row_num, column=5, value=amount)
        amt_cell.number_format = money_fmt

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"Initial file created: {OUTPUT}")

    # Verify totals
    print(f"Total rows: {len(rows)}")
    print(f"Grand total: {sum(r[4] for r in rows):.2f}")

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print("GUI_READY: launched LibreOffice Calc with DISPLAY=:0")

create_initial()
