"""
Reward Script: Create a pivot table grouping transaction dates by quarter,
               showing total sales per quarter per store.
Task ID: calc_pivot_034
Domain: libreoffice_calc
Scoring:
  C1 (0.25) - PivotTable sheet exists with correct header structure
  C2 (0.15) - Quarter labels Q1-Q4 present as row field
  C3 (0.25) - Known ground truth values match (Q1/Store1=28000, Q2/Store2=32000)
  C4 (0.20) - Grand Total = 520000
  C5 (0.15) - Internal consistency: row/column totals sum correctly
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_034'


def persist_app_state(domain):
    """Attempt to save any unsaved LibreOffice state."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # ----------------------------------------------------------------
    # Component 1: PivotTable sheet exists with correct header structure (0.25)
    # The initial file has only 'StoreSales'. The task asks the agent to
    # create a pivot table — we expect a new sheet with quarter/store layout.
    # ----------------------------------------------------------------
    try:
        pivot_sheet = None
        for sn in wb.sheetnames:
            if sn.lower() != 'storesales':
                # Check if this sheet has a pivot-like structure
                ws_candidate = wb[sn]
                if ws_candidate.max_row >= 3 and ws_candidate.max_column >= 3:
                    pivot_sheet = ws_candidate
                    break

        if pivot_sheet is None:
            print("FAIL: Component 1 — No pivot table sheet found (only StoreSales exists)")
        else:
            # Check headers contain store identifiers in columns
            header_row = []
            for c in range(1, pivot_sheet.max_column + 1):
                val = pivot_sheet.cell(row=1, column=c).value
                if val is not None:
                    header_row.append(str(val).strip())

            # We expect the header row to contain store names (Store1..Store4)
            store_headers = [h for h in header_row if 'store' in h.lower() or 'Store' in h]
            has_quarter_header = any('quarter' in h.lower() or h.lower() in ('q1', 'q2', 'q3', 'q4', '') for h in header_row)
            has_total_header = any('total' in h.lower() or 'grand' in h.lower() for h in header_row)

            if len(store_headers) >= 2 and (has_quarter_header or has_total_header):
                print(f"PASS: Component 1 — Pivot sheet '{pivot_sheet.title}' found with headers: {header_row} (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component 1 — Pivot sheet found but headers don't match expected structure: {header_row}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    if pivot_sheet is None:
        # Cannot proceed without pivot sheet
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    # ----------------------------------------------------------------
    # Component 2: Quarter labels Q1-Q4 present as row labels (0.15)
    # The task groups dates by quarter. We expect rows labeled Q1, Q2, Q3, Q4.
    # ----------------------------------------------------------------
    try:
        quarter_labels_found = set()
        expected_quarters = {'Q1', 'Q2', 'Q3', 'Q4'}
        for r in range(2, pivot_sheet.max_row + 1):
            val = pivot_sheet.cell(row=r, column=1).value
            if val is not None:
                val_str = str(val).strip().upper()
                if val_str in expected_quarters:
                    quarter_labels_found.add(val_str)

        if quarter_labels_found == expected_quarters:
            print(f"PASS: Component 2 — All four quarter labels found: {sorted(quarter_labels_found)} (0.15 pts)")
            total_score += 0.15
        elif len(quarter_labels_found) >= 2:
            partial = 0.15 * len(quarter_labels_found) / 4
            print(f"PARTIAL: Component 2 — Found {len(quarter_labels_found)}/4 quarters: {sorted(quarter_labels_found)} ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Expected Q1-Q4 row labels, found: {sorted(quarter_labels_found)}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Helper: Build a lookup of the pivot data for components 3-5
    # Map (quarter_label, store_header) -> value
    # ----------------------------------------------------------------
    pivot_data = {}
    header_map = {}  # col_index -> header_name
    try:
        for c in range(1, pivot_sheet.max_column + 1):
            val = pivot_sheet.cell(row=1, column=c).value
            if val is not None:
                header_map[c] = str(val).strip()

        for r in range(2, pivot_sheet.max_row + 1):
            row_label = pivot_sheet.cell(row=r, column=1).value
            if row_label is not None:
                row_label_str = str(row_label).strip()
                for c in range(2, pivot_sheet.max_column + 1):
                    if c in header_map:
                        cell_val = pivot_sheet.cell(row=r, column=c).value
                        pivot_data[(row_label_str, header_map[c])] = cell_val
    except Exception as e:
        print(f"WARN: Could not build pivot data map: {e}")

    # ----------------------------------------------------------------
    # Component 3: Ground truth values match (0.25)
    # From context: Q1/Store1=28000, Q2/Store2=32000
    # ----------------------------------------------------------------
    try:
        checks_passed = 0
        total_checks = 2

        # Check Q1/Store1 = 28000
        q1_s1 = pivot_data.get(('Q1', 'Store1'))
        if q1_s1 is not None:
            try:
                if abs(float(q1_s1) - 28000) < 1.0:
                    print(f"PASS: Component 3a — Q1/Store1 = {q1_s1} (expected 28000)")
                    checks_passed += 1
                else:
                    print(f"FAIL: Component 3a — Q1/Store1 = {q1_s1}, expected 28000")
            except (ValueError, TypeError):
                print(f"FAIL: Component 3a — Q1/Store1 value not numeric: {q1_s1}")
        else:
            print(f"FAIL: Component 3a — Q1/Store1 cell not found in pivot data")

        # Check Q2/Store2 = 32000
        q2_s2 = pivot_data.get(('Q2', 'Store2'))
        if q2_s2 is not None:
            try:
                if abs(float(q2_s2) - 32000) < 1.0:
                    print(f"PASS: Component 3b — Q2/Store2 = {q2_s2} (expected 32000)")
                    checks_passed += 1
                else:
                    print(f"FAIL: Component 3b — Q2/Store2 = {q2_s2}, expected 32000")
            except (ValueError, TypeError):
                print(f"FAIL: Component 3b — Q2/Store2 value not numeric: {q2_s2}")
        else:
            print(f"FAIL: Component 3b — Q2/Store2 cell not found in pivot data")

        c3_score = 0.25 * checks_passed / total_checks
        if c3_score > 0:
            print(f"PASS: Component 3 — {checks_passed}/{total_checks} ground truth values correct ({c3_score:.2f} pts)")
            total_score += c3_score
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ----------------------------------------------------------------
    # Component 4: Grand Total = 520000 (0.20)
    # The overall grand total across all quarters and stores should be 520000.
    # ----------------------------------------------------------------
    try:
        grand_total_found = False
        # Look for the grand total in the last row, last column or any cell
        # labeled "Grand Total"
        for r in range(2, pivot_sheet.max_row + 1):
            row_label = pivot_sheet.cell(row=r, column=1).value
            if row_label is not None and 'total' in str(row_label).lower():
                # Check the last data column for this row (Grand Total column)
                for c in range(pivot_sheet.max_column, 1, -1):
                    hdr = header_map.get(c, '')
                    if 'total' in hdr.lower() or 'grand' in hdr.lower():
                        gt_val = pivot_sheet.cell(row=r, column=c).value
                        if gt_val is not None:
                            try:
                                if abs(float(gt_val) - 520000) < 1.0:
                                    print(f"PASS: Component 4 — Grand Total = {gt_val} (expected 520000) (0.20 pts)")
                                    total_score += 0.20
                                    grand_total_found = True
                                else:
                                    print(f"FAIL: Component 4 — Grand Total = {gt_val}, expected 520000")
                                    grand_total_found = True
                            except (ValueError, TypeError):
                                print(f"FAIL: Component 4 — Grand Total not numeric: {gt_val}")
                                grand_total_found = True
                        break
                if grand_total_found:
                    break

        if not grand_total_found:
            # Fallback: try to find 520000 anywhere in the pivot table
            for r in range(2, pivot_sheet.max_row + 1):
                for c in range(2, pivot_sheet.max_column + 1):
                    val = pivot_sheet.cell(row=r, column=c).value
                    if val is not None:
                        try:
                            if abs(float(val) - 520000) < 1.0:
                                print(f"PASS: Component 4 — Grand Total value 520000 found at row {r}, col {c} (0.20 pts)")
                                total_score += 0.20
                                grand_total_found = True
                                break
                        except (ValueError, TypeError):
                            pass
                if grand_total_found:
                    break

            if not grand_total_found:
                print("FAIL: Component 4 — Grand Total value 520000 not found anywhere in pivot table")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ----------------------------------------------------------------
    # Component 5: Internal consistency — row totals sum correctly (0.15)
    # Each quarterly row's Grand Total should equal the sum of that row's
    # store values. Also, each store column total should equal the sum of
    # that store's quarterly values.
    # ----------------------------------------------------------------
    try:
        consistency_checks = 0
        total_consistency = 0

        # Find store columns (exclude first col=row labels, and Grand Total col)
        store_cols = []
        total_col = None
        for c in range(2, pivot_sheet.max_column + 1):
            hdr = header_map.get(c, '')
            if 'total' in hdr.lower() or 'grand' in hdr.lower():
                total_col = c
            elif hdr:
                store_cols.append(c)

        # Check each quarter row's total
        for r in range(2, pivot_sheet.max_row + 1):
            row_label = pivot_sheet.cell(row=r, column=1).value
            if row_label is None:
                continue
            row_label_str = str(row_label).strip().upper()
            if row_label_str in ('Q1', 'Q2', 'Q3', 'Q4'):
                row_sum = 0
                valid = True
                for c in store_cols:
                    val = pivot_sheet.cell(row=r, column=c).value
                    if val is not None:
                        try:
                            row_sum += float(val)
                        except (ValueError, TypeError):
                            valid = False
                    else:
                        valid = False

                if valid and total_col is not None:
                    row_total = pivot_sheet.cell(row=r, column=total_col).value
                    if row_total is not None:
                        total_consistency += 1
                        try:
                            if abs(float(row_total) - row_sum) < 1.0:
                                consistency_checks += 1
                        except (ValueError, TypeError):
                            pass

        if total_consistency > 0 and consistency_checks == total_consistency:
            print(f"PASS: Component 5 — All {consistency_checks} row totals are internally consistent (0.15 pts)")
            total_score += 0.15
        elif total_consistency > 0 and consistency_checks > 0:
            partial = 0.15 * consistency_checks / total_consistency
            print(f"PARTIAL: Component 5 — {consistency_checks}/{total_consistency} row totals consistent ({partial:.2f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 — Row totals are not internally consistent (checked {total_consistency} rows, {consistency_checks} passed)")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
