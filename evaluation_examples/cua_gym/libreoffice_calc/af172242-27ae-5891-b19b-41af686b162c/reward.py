"""
Reward Script: Extract CV paper authors and affiliations into spreadsheet
Task ID: osworld_multi_apps_pdf_author_extract_002
Domain: libreoffice_calc
Scoring:
  - Component 1: Headers 'Author' and 'Affiliation' in row 1 (0.20 pts)
  - Component 2: Exactly 5 data rows (one per paper) (0.20 pts)
  - Component 3: All 5 author names present (0.30 pts)
  - Component 4: All 5 affiliations match expected values (0.20 pts)
  - Component 5: Data sorted alphabetically by author name (0.10 pts)
  Total: 1.00
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'osworld_multi_apps_pdf_author_extract_002'

# Expected ground truth from the 5 CV conference PDF papers
EXPECTED_DATA = [
    ('Chen Wei', 'Peking University'),
    ('Emma Novak', 'ETH Zurich'),
    ('Hiroshi Nakamura', 'University of Tokyo'),
    ('James Mitchell', 'Stanford University'),
    ('Priya Sharma', 'Indian Institute of Technology Bombay'),
]

# Sorted order of expected authors
EXPECTED_AUTHORS_SORTED = sorted([row[0] for row in EXPECTED_DATA])


def normalize(s):
    """Normalize a string for comparison: strip whitespace, lowercase."""
    if s is None:
        return ''
    return str(s).strip().lower()


def verify_task(file_path):
    """
    Verify task completion: checks that cv_authors.xlsx was created with
    correct headers, 5 data rows (one per paper), proper author/affiliation
    values, and alphabetical sorting by author name.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition gate: load the file
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # --- Component 1: Correct headers 'Author' and 'Affiliation' in row 1 (0.20 pts) ---
    try:
        header_a = ws.cell(row=1, column=1).value
        header_b = ws.cell(row=1, column=2).value
        if normalize(header_a) == 'author' and normalize(header_b) == 'affiliation':
            print(f"PASS: Component 1 — Headers are correct: '{header_a}', '{header_b}' (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 1 — Expected headers 'Author', 'Affiliation', found: '{header_a}', '{header_b}'")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check headers: {e}")

    # --- Component 2: Exactly 5 data rows (0.20 pts) ---
    try:
        # Count non-empty rows after header row
        data_rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(cell is not None and str(cell).strip() != '' for cell in row):
                data_rows.append(row)
        row_count = len(data_rows)
        if row_count == 5:
            print(f"PASS: Component 2 — Found exactly 5 data rows (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 2 — Expected 5 data rows, found: {row_count}")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not count data rows: {e}")

    # --- Component 3: All 5 expected author names are present (0.30 pts) ---
    try:
        actual_authors = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            author = row[0] if len(row) > 0 else None
            if author is not None and str(author).strip() != '':
                actual_authors.append(str(author).strip())

        expected_authors_set = set(normalize(a) for a in EXPECTED_AUTHORS_SORTED)
        actual_authors_norm = set(normalize(a) for a in actual_authors)

        missing_authors = expected_authors_set - actual_authors_norm
        if not missing_authors:
            print(f"PASS: Component 3 — All 5 expected authors present: {actual_authors} (0.30 pts)")
            total_score += 0.30
        else:
            print(f"FAIL: Component 3 — Missing authors: {missing_authors}. Found: {actual_authors}")
    except Exception as e:
        print(f"ERROR: Component 3 — Could not check authors: {e}")

    # --- Component 4: All 5 affiliations match expected values (0.20 pts) ---
    try:
        # Build a mapping of author -> affiliation from the spreadsheet
        actual_mapping = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 2:
                author = row[0]
                affiliation = row[1]
                if author is not None and str(author).strip() != '':
                    actual_mapping[normalize(str(author).strip())] = str(affiliation).strip() if affiliation else ''

        expected_mapping = {normalize(a): aff for (a, aff) in EXPECTED_DATA}
        mismatches = []
        for exp_author_norm, exp_affil in expected_mapping.items():
            actual_affil = actual_mapping.get(exp_author_norm, None)
            if actual_affil is None or normalize(actual_affil) != normalize(exp_affil):
                mismatches.append(f"Author '{exp_author_norm}': expected '{exp_affil}', got '{actual_affil}'")

        if not mismatches:
            print(f"PASS: Component 4 — All 5 affiliations match expected values (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 4 — Affiliation mismatches: {mismatches}")
    except Exception as e:
        print(f"ERROR: Component 4 — Could not check affiliations: {e}")

    # --- Component 5: Data sorted alphabetically by author name (0.10 pts) ---
    try:
        actual_authors_in_order = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            author = row[0] if len(row) > 0 else None
            if author is not None and str(author).strip() != '':
                actual_authors_in_order.append(str(author).strip())

        if actual_authors_in_order == sorted(actual_authors_in_order):
            print(f"PASS: Component 5 — Data is sorted alphabetically by author name (0.10 pts)")
            print(f"  Order: {actual_authors_in_order}")
            total_score += 0.10
        else:
            print(f"FAIL: Component 5 — Data not sorted alphabetically. Found: {actual_authors_in_order}, expected: {sorted(actual_authors_in_order)}")
    except Exception as e:
        print(f"ERROR: Component 5 — Could not check sort order: {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score:.2f}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path in the given env
file_path = f'{WORKDIR}/cv_authors.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
