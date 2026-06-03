"""
Initial Setup: Task tracker spreadsheet WITHOUT dropdown validation on Priority column
Task ID: osworld_calc_data_validation_dropdown_003
Domain: libreoffice_calc

Creates a project task tracker with columns: Task ID, Task Name, Priority (empty),
Assignee, Due Date. The Priority column (C) has NO data validation — that is what
the agent must add.
"""

import os
import shlex
import subprocess
import time

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'osworld_calc_data_validation_dropdown_003'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Task Tracker ---
    ws = wb.active
    ws.title = "Task Tracker"

    # Headers
    headers = ["Task ID", "Task Name", "Priority", "Assignee", "Due Date"]
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_font = Font(name="Calibri", bold=True, color="FFFFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Realistic project task data (Priority column left EMPTY — no values)
    data = [
        ["TSK-001", "Design authentication module",          None, "Sarah Chen",      "2025-04-10"],
        ["TSK-002", "Implement database schema migration",   None, "Marcus Johnson",  "2025-04-15"],
        ["TSK-003", "Write unit tests for payment service",  None, "Priya Nair",      "2025-04-18"],
        ["TSK-004", "Update API documentation",              None, "Leon Fischer",    "2025-04-20"],
        ["TSK-005", "Fix login redirect bug",                None, "Sarah Chen",      "2025-04-08"],
        ["TSK-006", "Refactor notification service",         None, "Tomoko Yamada",   "2025-04-25"],
        ["TSK-007", "Deploy staging environment",            None, "Marcus Johnson",  "2025-04-12"],
        ["TSK-008", "Integrate third-party analytics SDK",   None, "Priya Nair",      "2025-04-30"],
        ["TSK-009", "Resolve memory leak in worker process", None, "Leon Fischer",    "2025-04-09"],
        ["TSK-010", "Create onboarding email templates",     None, "Tomoko Yamada",   "2025-05-02"],
        ["TSK-011", "Review security audit findings",        None, "Sarah Chen",      "2025-04-22"],
        ["TSK-012", "Set up CI/CD pipeline for mobile app",  None, "Marcus Johnson",  "2025-05-05"],
    ]

    for r, row_data in enumerate(data, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 14

    # Row height for header
    ws.row_dimensions[1].height = 22

    # Freeze header row
    ws.freeze_panes = "A2"

    # NOTE: NO DataValidation on column C — that is what the agent must add.

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup: open the file in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
