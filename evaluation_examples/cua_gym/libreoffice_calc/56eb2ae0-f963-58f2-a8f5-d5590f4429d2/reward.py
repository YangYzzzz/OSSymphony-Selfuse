"""
Reward Script: Protect specific cells in a sheet by unlocking C2:C50 then protecting the sheet.
Task ID: calc_gsi_033
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): Sheet protection is enabled
  Component 2 (0.5): Cells C2:C50 are unlocked (locked=False)
  Component 3 (0.2): Boundary cells outside C2:C50 remain locked
"""

import os

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_033'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb.active

    # Component 1: Sheet protection is enabled (0.3 points)
    # Initial env has protection=False; golden env has protection=True
    try:
        if ws.protection.sheet:
            print(f"PASS: Component 1 — Sheet protection is enabled (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — Sheet protection is NOT enabled (sheet={ws.protection.sheet})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Cells C2:C50 are unlocked (0.5 points)
    # Initial env has all cells locked=True; golden env has C2:C50 locked=False
    # Check a representative sample of cells in the range
    try:
        unlocked_count = 0
        total_cells = 49  # C2 through C50

        for row_num in range(2, 51):
            cell = ws.cell(row=row_num, column=3)  # Column C
            if not cell.protection.locked:
                unlocked_count += 1

        if unlocked_count == total_cells:
            print(f"PASS: Component 2 — All {total_cells} cells in C2:C50 are unlocked (0.5 pts)")
            total_score += 0.5
        elif unlocked_count > 0:
            # Partial credit: proportion of unlocked cells
            partial = 0.5 * (unlocked_count / total_cells)
            print(f"PARTIAL: Component 2 — {unlocked_count}/{total_cells} cells unlocked in C2:C50 ({partial:.3f} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No cells in C2:C50 are unlocked ({unlocked_count}/{total_cells})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Boundary cells outside C2:C50 remain locked (0.2 points)
    # This only matters if the task was actually attempted (sheet protection on + some unlocking done)
    # Check: C1 (header), C51 (just outside range), A2, B2, D2 (other columns)
    try:
        boundary_cells = ['C1', 'C51', 'A1', 'A2', 'B2', 'D2']
        all_locked = True
        failed_cells = []

        for coord in boundary_cells:
            cell = ws[coord]
            if not cell.protection.locked:
                all_locked = False
                failed_cells.append(coord)

        # Only award points if sheet protection is on AND C2:C50 were unlocked
        # This prevents awarding points on initial_env where everything is locked
        if ws.protection.sheet and unlocked_count > 0:
            if all_locked:
                print(f"PASS: Component 3 — Boundary cells {boundary_cells} remain locked (0.2 pts)")
                total_score += 0.2
            else:
                print(f"FAIL: Component 3 — Some boundary cells are unlocked: {failed_cells}")
        else:
            print(f"SKIP: Component 3 — Preconditions not met (protection={ws.protection.sheet}, unlocked_count={unlocked_count})")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state
persist_app_state("libreoffice_calc")

# Test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
