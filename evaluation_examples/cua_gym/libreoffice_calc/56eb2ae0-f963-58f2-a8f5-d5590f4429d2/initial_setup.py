"""
Initial Setup: Protect specific cells in a sheet by first unlocking the editable range C2:C50,
then protecting the sheet so only the unlocked cells can be edited.
Task ID: calc_gsi_033
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection

WORKDIR = '/home/user'
TASK_ID = 'calc_gsi_033'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Entry Form"

    # --- Styling ---
    header_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FF2F5496", end_color="FF2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    label_font = Font(name="Arial", size=11, bold=True, color="333333")
    formula_font = Font(name="Arial", size=11, italic=True, color="0070C0")
    data_font = Font(name="Arial", size=11)
    thin_border = Border(
        left=Side(style="thin", color="999999"),
        right=Side(style="thin", color="999999"),
        top=Side(style="thin", color="999999"),
        bottom=Side(style="thin", color="999999"),
    )
    light_gray_fill = PatternFill(start_color="FFF2F2F2", end_color="FFF2F2F2", fill_type="solid")

    # --- Headers (Row 1) ---
    headers = ["Field Name", "Reference Value", "Input Value", "Notes"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Column widths ---
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 35

    # --- Form data ---
    # Column A: labels, Column B: formulas/reference values, Column C: user input area, Column D: notes
    form_rows = [
        ("Employee ID", "EMP-2025-001", "EMP-2025-001", "Format: EMP-YYYY-NNN"),
        ("Full Name", None, "Sarah Chen", "First and last name"),
        ("Department", "Engineering", "Engineering", "Must match department list"),
        ("Job Title", None, "Senior Software Engineer", "Current position"),
        ("Start Date", "2023-01-15", "2023-01-15", "Format: YYYY-MM-DD"),
        ("Annual Salary", 92500, 92500, "Base salary in USD"),
        ("Bonus Percentage", 0.15, 0.15, "Decimal format (0.15 = 15%)"),
        ("Manager Name", "David Park", "David Park", "Direct supervisor"),
        ("Office Location", "Building A, Floor 3", "Building A, Floor 3", "Building and floor"),
        ("Email Address", None, "sarah.chen@company.com", "Corporate email"),
        ("Phone Extension", 4521, 4521, "4-digit extension"),
        ("Emergency Contact", None, "Michael Chen", "Name of contact"),
        ("Emergency Phone", None, "+1-555-0142", "Include country code"),
        ("Health Plan", "Premium Plus", "Premium Plus", "Select from plan options"),
        ("401k Contribution", 0.08, 0.08, "Percentage of salary"),
        ("PTO Balance (Days)", 18, 18, "Current balance"),
        ("Last Review Date", "2024-09-15", "2024-09-15", "Most recent performance review"),
        ("Review Score", 4.2, 4.2, "Scale: 1.0 to 5.0"),
        ("Training Hours YTD", 24, 24, "Year-to-date training"),
        ("Certification", None, "AWS Solutions Architect", "Professional certification"),
        ("Parking Permit", "P-0312", "P-0312", "Permit number"),
        ("Equipment Issued", "Laptop, Monitor, Keyboard", "Laptop, Monitor, Keyboard", "Company property"),
        ("Security Clearance", "Level 2", "Level 2", "Current clearance level"),
        ("Remote Work Days", 3, 3, "Days per week"),
        ("Notes", None, "Transferred from NYC office in Q2 2024", "Additional information"),
    ]

    for r, (label, ref_val, input_val, note) in enumerate(form_rows, 2):
        # Column A: Field labels
        cell_a = ws.cell(row=r, column=1, value=label)
        cell_a.font = label_font
        cell_a.border = thin_border
        if r % 2 == 0:
            cell_a.fill = light_gray_fill

        # Column B: Reference / formula values
        cell_b = ws.cell(row=r, column=2, value=ref_val)
        cell_b.font = formula_font
        cell_b.border = thin_border
        if r % 2 == 0:
            cell_b.fill = light_gray_fill

        # Column C: Input values (the editable column)
        cell_c = ws.cell(row=r, column=3, value=input_val)
        cell_c.font = data_font
        cell_c.border = thin_border
        if r % 2 == 0:
            cell_c.fill = light_gray_fill

        # Column D: Notes
        cell_d = ws.cell(row=r, column=4, value=note)
        cell_d.font = Font(name="Arial", size=10, color="666666")
        cell_d.border = thin_border
        if r % 2 == 0:
            cell_d.fill = light_gray_fill

    # Add some formulas in Column B that reference Column C
    ws.cell(row=8, column=2, value="=C6*C7")  # Bonus Amount = Salary * Bonus %
    ws.cell(row=8, column=2).font = formula_font
    ws.cell(row=2, column=2, value="=C2")  # Echo employee ID
    ws.cell(row=2, column=2).font = formula_font

    # Freeze top row
    ws.freeze_panes = "A2"

    # NOTE: All cells have default protection (locked=True) since we do NOT
    # unlock anything or protect the sheet. The task asks the agent to:
    # 1. Unlock C2:C50
    # 2. Then protect the sheet

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
