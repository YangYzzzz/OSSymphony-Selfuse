"""
Reward Script: Legal Timesheet Setup — ROUND hours, VLOOKUP rates, billing amounts, CF, and BillingSummary
Task ID: calc_gen_legal_024
Domain: libreoffice_calc
Scoring:
  - Component 1: G2:G301 contain ROUND(Fn,1) formulas           (0.25 pts)
  - Component 2: H2:H301 contain VLOOKUP to Rates sheet         (0.25 pts)
  - Component 3: I2:I301 contain Gn*Hn billing amount formulas  (0.20 pts)
  - Component 4: Conditional formatting on F2:F301 yellow > 12  (0.15 pts)
  - Component 5: BillingSummary sheet populated with SUMPRODUCT  (0.15 pts)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_legal_024'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Required sheets must exist
    required_sheets = ['TimeEntries', 'Rates', 'BillingSummary']
    for sheet in required_sheets:
        if sheet not in wb.sheetnames:
            print(f"CRITICAL: Required sheet '{sheet}' not found. Cannot score.")
            print("REWARD: 0.0")
            return 0.0

    ws_te = wb['TimeEntries']
    ws_bs = wb['BillingSummary']

    # Component 1: G2:G301 contain ROUND(Fn,1) formulas (0.25 points)
    # This FAILS on initial (G is empty) and PASSES on golden
    try:
        round_count = 0
        round_correct = 0
        for row in range(2, 302):
            g_val = ws_te.cell(row=row, column=7).value
            if g_val is not None:
                round_count += 1
                # Check it's a ROUND formula referencing F column
                if (isinstance(g_val, str) and
                        'ROUND' in g_val.upper() and
                        f'F{row}' in g_val):
                    round_correct += 1

        if round_correct >= 295:  # Allow small tolerance
            print(f"PASS: Component 1 — G2:G301 ROUND formulas present ({round_correct}/300 rows correct) (0.25 pts)")
            total_score += 0.25
        elif round_correct > 0:
            partial = round(0.25 * round_correct / 300, 4)
            print(f"PARTIAL: Component 1 — G column partial ({round_correct}/300 rows correct) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — G2:G301 ROUND formulas not found (found {round_count} non-empty, {round_correct} correct)")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: H2:H301 contain VLOOKUP to Rates sheet (0.25 points)
    # This FAILS on initial (H is empty) and PASSES on golden
    try:
        vlookup_count = 0
        vlookup_correct = 0
        for row in range(2, 302):
            h_val = ws_te.cell(row=row, column=8).value
            if h_val is not None:
                vlookup_count += 1
                # Check it's a VLOOKUP formula referencing Rates sheet
                if (isinstance(h_val, str) and
                        'VLOOKUP' in h_val.upper() and
                        'RATES' in h_val.upper()):
                    vlookup_correct += 1

        if vlookup_correct >= 295:  # Allow small tolerance
            print(f"PASS: Component 2 — H2:H301 VLOOKUP formulas present ({vlookup_correct}/300 rows correct) (0.25 pts)")
            total_score += 0.25
        elif vlookup_correct > 0:
            partial = round(0.25 * vlookup_correct / 300, 4)
            print(f"PARTIAL: Component 2 — H column partial ({vlookup_correct}/300 rows correct) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — H2:H301 VLOOKUP formulas not found (found {vlookup_count} non-empty, {vlookup_correct} correct)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: I2:I301 contain billing amount formulas Gn*Hn (0.20 points)
    # This FAILS on initial (I is empty) and PASSES on golden
    try:
        amount_count = 0
        amount_correct = 0
        for row in range(2, 302):
            i_val = ws_te.cell(row=row, column=9).value
            if i_val is not None:
                amount_count += 1
                # Check it's a formula multiplying G and H columns
                if (isinstance(i_val, str) and
                        f'G{row}' in i_val and
                        f'H{row}' in i_val and
                        '*' in i_val):
                    amount_correct += 1

        if amount_correct >= 295:  # Allow small tolerance
            print(f"PASS: Component 3 — I2:I301 billing amount formulas (G*H) present ({amount_correct}/300 rows) (0.20 pts)")
            total_score += 0.20
        elif amount_correct > 0:
            partial = round(0.20 * amount_correct / 300, 4)
            print(f"PARTIAL: Component 3 — I column partial ({amount_correct}/300 rows correct) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — I2:I301 billing formulas not found (found {amount_count} non-empty, {amount_correct} correct)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Conditional formatting on F2:F301 — yellow fill when F > 12 (0.15 points)
    # This FAILS on initial (no CF rules) and PASSES on golden
    try:
        cf_on_f_count = 0          # count of CF rules targeting F column
        gt12_rule_count = 0        # count of greaterThan 12 rules
        yellow_fill_count = 0      # count of rules with yellow fill

        for cf_range, cf_list in ws_te.conditional_formatting._cf_rules.items():
            cf_range_str = str(cf_range)
            # Check the range covers F column rows 2-301
            if 'F' in cf_range_str:
                for rule in cf_list:
                    cf_on_f_count += 1
                    # Check for greaterThan 12 condition
                    if (getattr(rule, 'type', None) == 'cellIs' and
                            getattr(rule, 'operator', None) == 'greaterThan' and
                            getattr(rule, 'formula', None) and
                            '12' in str(rule.formula)):
                        gt12_rule_count += 1
                    # Check for yellow fill
                    try:
                        fill_color = rule.dxf.fill.fgColor.rgb
                        # Yellow: FFFFFF00 or similar yellow
                        if 'FFFF00' in fill_color or fill_color in ('FFFFFF00', 'FFFF0000'):
                            yellow_fill_count += 1
                    except Exception:
                        pass

        if cf_on_f_count > 0 and gt12_rule_count > 0 and yellow_fill_count > 0:
            print(f"PASS: Component 4 — Conditional formatting on F column: yellow fill for > 12 hours (0.15 pts)")
            total_score += 0.15
        elif cf_on_f_count > 0 and gt12_rule_count > 0:
            print(f"PARTIAL: Component 4 — CF rule for > 12 found but yellow fill not confirmed (0.08 pts)")
            total_score += 0.08
        elif cf_on_f_count > 0:
            print(f"PARTIAL: Component 4 — CF on F column found but wrong condition (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 4 — No conditional formatting found on F column (hours > 12 yellow fill)")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # Component 5: BillingSummary sheet populated with SUMPRODUCT formulas (0.15 points)
    # This FAILS on initial (BillingSummary is empty) and PASSES on golden
    try:
        bs_max_row = ws_bs.max_row
        bs_max_col = ws_bs.max_column

        # Check headers
        header_a = ws_bs.cell(row=1, column=1).value
        header_b = ws_bs.cell(row=1, column=2).value
        header_c = ws_bs.cell(row=1, column=3).value

        # Count SUMPRODUCT formulas in column C
        sumproduct_count = 0
        for row in range(2, bs_max_row + 1):
            c_val = ws_bs.cell(row=row, column=3).value
            if (isinstance(c_val, str) and
                    'SUMPRODUCT' in c_val.upper() and
                    'TIMEENTRIES' in c_val.upper()):
                sumproduct_count += 1

        # Check structure: has client (A), month (B), sumproduct (C)
        has_headers = (header_a is not None and header_b is not None and header_c is not None)
        has_data = bs_max_row > 1 and sumproduct_count >= 10

        if has_headers and sumproduct_count >= 90:
            print(f"PASS: Component 5 — BillingSummary has headers and {sumproduct_count} SUMPRODUCT formulas (0.15 pts)")
            total_score += 0.15
        elif has_data:
            partial = round(0.15 * min(sumproduct_count / 96, 1.0), 4)
            print(f"PARTIAL: Component 5 — BillingSummary has {sumproduct_count} SUMPRODUCT formulas ({partial} pts)")
            total_score += partial
        elif bs_max_row > 1:
            print(f"PARTIAL: Component 5 — BillingSummary has data rows ({bs_max_row} rows) but no SUMPRODUCT formulas (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 5 — BillingSummary is empty (max_row={bs_max_row}, SUMPRODUCT count={sumproduct_count})")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
