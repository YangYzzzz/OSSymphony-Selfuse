"""
Initial Setup: Legal timesheet for billing purposes
Task ID: calc_gen_legal_024
Domain: libreoffice_calc
"""

import os
import random
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gen_legal_024'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

# 7 attorneys with realistic names
ATTORNEYS = [
    'Alexandra Rivera',
    'Benjamin Hartley',
    'Catherine Moss',
    'Daniel Whitfield',
    'Elena Suzuki',
    'Franklin Okafor',
    'Grace Thornton',
]

# Clients
CLIENTS = [
    'Meridian Corp',
    'Blackstone Industries',
    'Fairway Holdings',
    'Crestwood Technologies',
    'Northgate Partners',
    'Summit Ventures',
    'Harborview Medical',
    'Apex Logistics',
]

# Matter types
MATTERS = [
    'Contract Review',
    'M&A Due Diligence',
    'Litigation Defense',
    'Regulatory Compliance',
    'IP Prosecution',
    'Employment Dispute',
    'Real Estate Acquisition',
    'Corporate Governance',
    'Securities Filing',
    'Bankruptcy Proceedings',
]

# Notes
NOTES_OPTIONS = [
    'Client meeting',
    'Document review',
    'Research and analysis',
    'Drafted motion',
    'Conference call',
    'Deposition preparation',
    'Court filing',
    'Negotiation session',
    'Internal strategy meeting',
    'Due diligence review',
    '',
    '',
    '',
]

random.seed(42)

def random_hours():
    """Generate random hours in 0.1 increments, mostly 0.1-8.0, occasionally > 12."""
    r = random.random()
    if r < 0.05:
        # ~5% over 12 hours (erroneous)
        return round(random.uniform(12.1, 16.0), 2)
    elif r < 0.15:
        # 10% longer days
        return round(random.uniform(8.1, 12.0), 2)
    else:
        return round(random.uniform(0.1, 8.0), 2)

def random_date():
    """Random date in 2024."""
    start = datetime.date(2024, 1, 1)
    end = datetime.date(2024, 12, 31)
    delta = (end - start).days
    return start + datetime.timedelta(days=random.randint(0, delta))

def create_initial():
    wb = openpyxl.Workbook()

    # ─── Sheet 1: TimeEntries ───
    ws1 = wb.active
    ws1.title = 'TimeEntries'

    headers = ['Entry ID', 'Attorney', 'Client', 'Date', 'Matter',
               'Hours', 'Billable Hours', 'Rate', 'Amount', 'Notes']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color='FFD9E1F2', end_color='FFD9E1F2', fill_type='solid')

    # 300 time entries (G=Billable Hours, H=Rate, I=Amount intentionally empty)
    for i in range(300):
        row = i + 2
        entry_id = f'TE-{1001 + i}'
        attorney = random.choice(ATTORNEYS)
        client = random.choice(CLIENTS)
        date_val = random_date()
        matter = random.choice(MATTERS)
        hours = random_hours()
        note = random.choice(NOTES_OPTIONS)

        ws1.cell(row=row, column=1, value=entry_id)
        ws1.cell(row=row, column=2, value=attorney)
        ws1.cell(row=row, column=3, value=client)
        date_cell = ws1.cell(row=row, column=4, value=date_val)
        date_cell.number_format = 'yyyy-mm-dd'
        ws1.cell(row=row, column=5, value=matter)
        ws1.cell(row=row, column=6, value=hours)
        # Columns G (7), H (8), I (9) intentionally left empty
        ws1.cell(row=row, column=10, value=note)

    # Set column widths for readability
    ws1.column_dimensions['A'].width = 10
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 12
    ws1.column_dimensions['E'].width = 24
    ws1.column_dimensions['F'].width = 8
    ws1.column_dimensions['G'].width = 14
    ws1.column_dimensions['H'].width = 8
    ws1.column_dimensions['I'].width = 12
    ws1.column_dimensions['J'].width = 24

    ws1.freeze_panes = 'A2'

    # ─── Sheet 2: Rates ───
    ws2 = wb.create_sheet('Rates')
    ws2.cell(row=1, column=1, value='Attorney').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Hourly Rate').font = Font(bold=True)

    # 7 attorneys with hourly rates $250 to $650
    rate_data = [
        ('Alexandra Rivera', 450),
        ('Benjamin Hartley', 350),
        ('Catherine Moss',   550),
        ('Daniel Whitfield', 650),
        ('Elena Suzuki',     400),
        ('Franklin Okafor',  250),
        ('Grace Thornton',   500),
    ]
    for r, (atty, rate) in enumerate(rate_data, 2):
        ws2.cell(row=r, column=1, value=atty)
        rate_cell = ws2.cell(row=r, column=2, value=rate)
        rate_cell.number_format = '$#,##0'

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 14

    # ─── Sheet 3: BillingSummary (empty) ───
    ws3 = wb.create_sheet('BillingSummary')
    # Intentionally left empty — task requires agent to populate it

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  TimeEntries: 300 rows with Hours (raw), G/H/I empty')
    print(f'  Rates: 7 attorneys, rates $250–$650')
    print(f'  BillingSummary: empty (to be filled by agent)')

create_initial()
