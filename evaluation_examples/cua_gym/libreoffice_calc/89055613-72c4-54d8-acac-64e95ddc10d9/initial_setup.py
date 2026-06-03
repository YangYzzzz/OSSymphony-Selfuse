"""
Initial Setup: Print area is A1:D20, data extends to H50, comments on several cells.
Task ID: calc_mcp_095
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_095'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = Workbook()
    ws = wb.active
    ws.title = 'Print Preview'

    # --- Headers (row 1) ---
    headers = [
        'Employee ID', 'Full Name', 'Department', 'Position',
        'Base Salary', 'Bonus', 'Tax Rate', 'Net Pay'
    ]
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='FF2F5496', end_color='FF2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    thin = Side(style='thin', color='000000')
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data rows 2-50 (49 rows of employee data) ---
    departments = ['Engineering', 'Marketing', 'Finance', 'Operations', 'HR', 'Sales', 'Legal', 'IT Support']
    positions = {
        'Engineering': ['Software Engineer', 'Senior Developer', 'Tech Lead', 'DevOps Engineer', 'QA Analyst'],
        'Marketing': ['Marketing Analyst', 'Content Strategist', 'SEO Specialist', 'Brand Manager', 'Campaign Lead'],
        'Finance': ['Financial Analyst', 'Accountant', 'Controller', 'Auditor', 'Budget Planner'],
        'Operations': ['Operations Manager', 'Supply Chain Analyst', 'Logistics Coordinator', 'Process Engineer', 'Facilities Lead'],
        'HR': ['HR Specialist', 'Recruiter', 'Training Coordinator', 'Benefits Analyst', 'HR Manager'],
        'Sales': ['Sales Representative', 'Account Executive', 'Sales Manager', 'Business Development', 'Territory Manager'],
        'Legal': ['Paralegal', 'Legal Counsel', 'Compliance Officer', 'Contract Specialist', 'IP Attorney'],
        'IT Support': ['Help Desk Analyst', 'System Administrator', 'Network Engineer', 'Security Analyst', 'IT Manager'],
    }
    names = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien', 'Yuki Tanaka',
        'Elena Rodriguez', 'David Kim', 'Fatima Al-Hassan', 'Robert Williams', 'Anika Sharma',
        'Thomas Mueller', 'Grace Okonkwo', 'Lucas Fernandez', 'Mei Lin', 'Ahmed Khalil',
        'Sophie Laurent', 'Raj Gupta', 'Isabella Costa', 'William Chang', 'Nadia Petrova',
        'Carlos Reyes', 'Hannah Wright', 'Omar Farouk', 'Lena Johansson', 'Derek Foster',
        'Amara Diallo', 'Kevin Park', 'Zara Hussain', 'Michael Brown', 'Yuna Sato',
        'Benjamin Hart', 'Chloe Nguyen', 'Daniel Torres', 'Eva Kovacs', 'Felix Andersen',
        'Georgia Papadopoulos', 'Hassan Mirza', 'Iris Nakamura', 'Jack Morrison', 'Kira Volkov',
        'Leonardo Bianchi', 'Maya Singh', 'Nathan Reed', 'Olivia Campbell', 'Patrick Dubois',
        'Quinn Bailey', 'Rosa Mendez', 'Samuel Ito', 'Tara O\'Connor',
    ]

    import random
    random.seed(42)

    for i in range(49):
        row = i + 2
        emp_id = f'EMP-{1000 + i}'
        name = names[i]
        dept = departments[i % len(departments)]
        pos = random.choice(positions[dept])
        base_salary = random.randint(55, 135) * 1000
        bonus = round(base_salary * random.uniform(0.03, 0.15), 2)
        tax_rate = round(random.uniform(0.18, 0.35), 4)
        net_pay = round((base_salary + bonus) * (1 - tax_rate), 2)

        ws.cell(row=row, column=1, value=emp_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=dept)
        ws.cell(row=row, column=4, value=pos)
        ws.cell(row=row, column=5, value=base_salary)
        ws.cell(row=row, column=6, value=bonus)
        ws.cell(row=row, column=7, value=tax_rate)
        ws.cell(row=row, column=8, value=net_pay)

        # Format salary/bonus/net pay columns
        ws.cell(row=row, column=5).number_format = '$#,##0'
        ws.cell(row=row, column=6).number_format = '$#,##0.00'
        ws.cell(row=row, column=7).number_format = '0.00%'
        ws.cell(row=row, column=8).number_format = '$#,##0.00'

    # --- Set print area to A1:D20 (NOT A1:H50 -- that's the task!) ---
    ws.print_area = 'A1:D20'

    # --- Add comments/notes to several cells ---
    ws['A1'].comment = Comment('Primary key for employee records', 'Admin')
    ws['B5'].comment = Comment('Promoted to Senior level in Q2 2025', 'HR Department')
    ws['C10'].comment = Comment('Department restructuring planned for Q3', 'Management')
    ws['E1'].comment = Comment('Base salary before deductions and bonuses', 'Finance Team')
    ws['F15'].comment = Comment('Bonus includes performance and retention components', 'Compensation')
    ws['H20'].comment = Comment('Net pay after all deductions applied', 'Payroll')
    ws['D30'].comment = Comment('Position title updated per new job classification system', 'HR Department')
    ws['G40'].comment = Comment('Tax rate adjusted per 2025 federal guidelines', 'Tax Compliance')

    # cellComments is left at default (None = do not print) -- task asks to set it

    # --- Column widths ---
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 14

    # --- Freeze header row ---
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
