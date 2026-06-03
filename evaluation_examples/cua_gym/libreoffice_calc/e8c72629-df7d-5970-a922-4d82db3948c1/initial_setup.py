"""
Initial Setup: Parent Contact List for Progress Report Mailing
Task ID: calc_edu_parent_contact_merge_032
Domain: libreoffice_calc

Creates a spreadsheet with 80 student contact records.
Sheet 'Contacts' has headers: Last Name, First Name, Phone, Email, Student GPA, Full Name, Formatted Phone, Grade Status
Columns F (Full Name), G (Formatted Phone), H (Grade Status) are empty — to be filled by agent.
"""

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_parent_contact_merge_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'

def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'Contacts'

    # Headers
    headers = ['Last Name', 'First Name', 'Phone', 'Email', 'Student GPA', 'Full Name', 'Formatted Phone', 'Grade Status']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # 80 realistic student records
    # Phone numbers stored as text (10 digits, no formatting)
    students = [
        ('Anderson', 'Emily',    '5551234567', 'emily.anderson@school.edu',   3.8),
        ('Baxter',   'James',    '5552345678', 'james.baxter@school.edu',     2.4),
        ('Chen',     'Sophia',   '5553456789', 'sophia.chen@school.edu',      3.6),
        ('Davis',    'Marcus',   '5554567890', 'marcus.davis@school.edu',     1.7),
        ('Evans',    'Olivia',   '5555678901', 'olivia.evans@school.edu',     2.9),
        ('Flores',   'Noah',     '5556789012', 'noah.flores@school.edu',      3.5),
        ('Garcia',   'Ava',      '5557890123', 'ava.garcia@school.edu',       0.8),
        ('Harris',   'Ethan',    '5558901234', 'ethan.harris@school.edu',     3.1),
        ('Ingram',   'Isabella', '5559012345', 'isabella.ingram@school.edu',  2.0),
        ('Jackson',  'Liam',     '5550123456', 'liam.jackson@school.edu',     3.9),
        ('Kim',      'Mia',      '5551357924', 'mia.kim@school.edu',          1.3),
        ('Lopez',    'William',  '5552468013', 'william.lopez@school.edu',    2.7),
        ('Martinez', 'Charlotte','5553579024', 'charlotte.martinez@school.edu',3.4),
        ('Nelson',   'Benjamin', '5554680135', 'benjamin.nelson@school.edu',  0.5),
        ('Ortiz',    'Amelia',   '5555791246', 'amelia.ortiz@school.edu',     2.2),
        ('Parker',   'Lucas',    '5556802357', 'lucas.parker@school.edu',     3.7),
        ('Quinn',    'Harper',   '5557913468', 'harper.quinn@school.edu',     1.9),
        ('Ramirez',  'Alexander','5558024579', 'alexander.ramirez@school.edu',3.3),
        ('Sanders',  'Evelyn',   '5559135680', 'evelyn.sanders@school.edu',   2.6),
        ('Torres',   'Henry',    '5550246791', 'henry.torres@school.edu',     3.6),
        ('Upton',    'Abigail',  '5551357802', 'abigail.upton@school.edu',    0.9),
        ('Vargas',   'Sebastian','5552468913', 'sebastian.vargas@school.edu', 2.8),
        ('Walker',   'Scarlett', '5553579024', 'scarlett.walker@school.edu',  3.5),
        ('Xavier',   'Michael',  '5554680135', 'michael.xavier@school.edu',   1.4),
        ('Young',    'Aria',     '5555791246', 'aria.young@school.edu',       2.3),
        ('Zhang',    'Daniel',   '5556802357', 'daniel.zhang@school.edu',     3.8),
        ('Adams',    'Grace',    '5557913468', 'grace.adams@school.edu',      2.1),
        ('Brooks',   'Owen',     '5558024579', 'owen.brooks@school.edu',      3.0),
        ('Carter',   'Chloe',    '5559135680', 'chloe.carter@school.edu',     1.6),
        ('Dixon',    'Julian',   '5550246791', 'julian.dixon@school.edu',     3.4),
        ('Edwards',  'Penelope', '5551368024', 'penelope.edwards@school.edu', 0.7),
        ('Fisher',   'Jack',     '5552479135', 'jack.fisher@school.edu',      2.5),
        ('Grant',    'Lily',     '5553580246', 'lily.grant@school.edu',       3.6),
        ('Hayes',    'Elijah',   '5554691357', 'elijah.hayes@school.edu',     1.1),
        ('Irwin',    'Zoey',     '5555702468', 'zoey.irwin@school.edu',       2.9),
        ('Jensen',   'Aiden',    '5556813579', 'aiden.jensen@school.edu',     3.7),
        ('Kelly',    'Natalie',  '5557924680', 'natalie.kelly@school.edu',    0.4),
        ('Lewis',    'Caleb',    '5558035791', 'caleb.lewis@school.edu',      2.6),
        ('Morgan',   'Addison',  '5559146802', 'addison.morgan@school.edu',   3.2),
        ('Nash',     'Wyatt',    '5550257913', 'wyatt.nash@school.edu',       1.8),
        ('Oliver',   'Layla',    '5551368024', 'layla.oliver@school.edu',     3.5),
        ('Price',    'Jayden',   '5552479135', 'jayden.price@school.edu',     2.4),
        ('Reed',     'Hannah',   '5553580246', 'hannah.reed@school.edu',      3.9),
        ('Scott',    'Gabriel',  '5554691357', 'gabriel.scott@school.edu',    1.5),
        ('Taylor',   'Aubrey',   '5555702468', 'aubrey.taylor@school.edu',    2.7),
        ('Underwood','Ryan',     '5556813579', 'ryan.underwood@school.edu',   3.3),
        ('Vincent',  'Brooklyn', '5557924680', 'brooklyn.vincent@school.edu', 0.6),
        ('Warren',   'Christian','5558035791', 'christian.warren@school.edu', 2.1),
        ('Xu',       'Nora',     '5559146802', 'nora.xu@school.edu',          3.8),
        ('York',     'Eli',      '5550257913', 'eli.york@school.edu',         1.2),
        ('Zimmerman','Claire',   '5551479024', 'claire.zimmerman@school.edu', 2.8),
        ('Abbott',   'Isaac',    '5552580135', 'isaac.abbott@school.edu',     3.6),
        ('Bishop',   'Savannah', '5553691246', 'savannah.bishop@school.edu',  0.3),
        ('Cannon',   'Landon',   '5554702357', 'landon.cannon@school.edu',    2.2),
        ('Dean',     'Samantha', '5555813468', 'samantha.dean@school.edu',    3.4),
        ('Elliott',  'Jordan',   '5556924579', 'jordan.elliott@school.edu',   1.7),
        ('Ford',     'Leah',     '5557035680', 'leah.ford@school.edu',        2.9),
        ('Gomez',    'Evan',     '5558146791', 'evan.gomez@school.edu',       3.7),
        ('Hood',     'Violet',   '5559257802', 'violet.hood@school.edu',      1.0),
        ('Ingles',   'Connor',   '5550368913', 'connor.ingles@school.edu',    2.5),
        ('Jacobs',   'Stella',   '5551479024', 'stella.jacobs@school.edu',    3.3),
        ('Knox',     'Dominic',  '5552580135', 'dominic.knox@school.edu',     0.9),
        ('Lane',     'Lucy',     '5553691246', 'lucy.lane@school.edu',        2.6),
        ('Marsh',    'Aaron',    '5554702357', 'aaron.marsh@school.edu',      3.8),
        ('Norton',   'Bella',    '5555813468', 'bella.norton@school.edu',     1.4),
        ('Owen',     'Isaiah',   '5556924579', 'isaiah.owen@school.edu',      2.3),
        ('Penn',     'Madeline', '5557035680', 'madeline.penn@school.edu',    3.6),
        ('Ross',     'Xavier',   '5558146791', 'xavier.ross@school.edu',      0.7),
        ('Shaw',     'Autumn',   '5559257802', 'autumn.shaw@school.edu',      2.7),
        ('Stone',    'Adrian',   '5550368913', 'adrian.stone@school.edu',     3.5),
        ('Tran',     'Jasmine',  '5551480024', 'jasmine.tran@school.edu',     1.6),
        ('Ulrich',   'Anthony',  '5552591135', 'anthony.ulrich@school.edu',   2.4),
        ('Vance',    'Piper',    '5553602246', 'piper.vance@school.edu',      3.9),
        ('Webb',     'Brandon',  '5554713357', 'brandon.webb@school.edu',     1.1),
        ('Xu',       'Diana',    '5555824468', 'diana.xu2@school.edu',        2.8),
        ('Yates',    'Tyler',    '5556935579', 'tyler.yates@school.edu',      3.2),
        ('Zhu',      'Julia',    '5558046680', 'julia.zhu@school.edu',        0.5),
        ('Alston',   'Rachel',   '5557158791', 'rachel.alston@school.edu',    2.0),
        ('Burns',    'Marcus',   '5558269802', 'marcus.burns@school.edu',     3.5),
        ('Cole',     'Freya',    '5559370913', 'freya.cole@school.edu',       1.3),
    ]

    for r, (last, first, phone, email, gpa) in enumerate(students, 2):
        ws.cell(row=r, column=1, value=last)
        ws.cell(row=r, column=2, value=first)
        # Phone stored as text (10-digit, no formatting)
        ws.cell(row=r, column=3, value=phone)
        ws.cell(row=r, column=4, value=email)
        ws.cell(row=r, column=5, value=gpa)
        # Columns F, G, H are intentionally left empty

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Sheet: Contacts')
    print(f'  Rows: {len(students)} student records (rows 2-{len(students)+1})')
    assert len(students) == 80, f'Expected 80 students, got {len(students)}'
    print(f'  Columns F, G, H are empty (to be filled by agent)')

create_initial()
