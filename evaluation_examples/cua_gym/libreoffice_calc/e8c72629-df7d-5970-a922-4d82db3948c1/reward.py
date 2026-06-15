"""
Reward Script: Prepare parent contact list with Full Name, Formatted Phone, and Grade Status
Task ID: calc_edu_parent_contact_merge_032
Domain: libreoffice_calc
Scoring:
  - Component 1: Full Name formulas in F2:F81 (0.35 pts) - =B{n}&" "&A{n} pattern
  - Component 2: Formatted Phone formulas in G2:G81 (0.35 pts) - (XXX) XXX-XXXX pattern
  - Component 3: Grade Status IFS formulas in H2:H81 (0.30 pts) - IFS with 4 thresholds
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_edu_parent_contact_merge_032'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Add formulas for Full Name (F), Formatted Phone (G), and Grade Status (H)
    for all 80 student rows (rows 2-81) in the 'Contacts' sheet.
    Initial file has F, G, H all empty — scoring only measures what was added.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Contacts' sheet must exist
    if 'Contacts' not in wb.sheetnames:
        print("FAIL: 'Contacts' sheet not found in workbook")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Contacts']

    # Precondition: sheet must have at least 81 rows (header + 80 students)
    if ws.max_row < 81:
        print(f"FAIL: Expected at least 81 rows, found {ws.max_row}")
        print("REWARD: 0.0")
        return 0.0

    # -------------------------------------------------------------------
    # Component 1: Full Name column F (0.35 points)
    # Each row n should have formula: =B{n}&" "&A{n}
    # This was empty in the initial file — verifies task change.
    # We check that at least 75/80 rows have a formula matching the pattern.
    # -------------------------------------------------------------------
    try:
        full_name_correct = 0
        full_name_total = 0
        full_name_pattern = re.compile(r'^=B(\d+)&" "&A\1$', re.IGNORECASE)

        for row in range(2, 82):
            val = ws.cell(row=row, column=6).value  # Column F
            full_name_total += 1
            if val is not None and isinstance(val, str):
                # Normalize whitespace for comparison
                val_norm = val.strip().replace(' ', ' ')
                # Check pattern: =B{n}&" "&A{n} where n == current row
                m = full_name_pattern.match(val_norm)
                if m and int(m.group(1)) == row:
                    full_name_correct += 1
                else:
                    # Also accept case-insensitive variants
                    expected = f'=B{row}&" "&A{row}'
                    if val.upper().replace(' ', '') == expected.upper().replace(' ', ''):
                        full_name_correct += 1

        if full_name_correct == 80:
            print(f"PASS: Component 1 — Full Name formulas: all 80/80 rows correct (0.35 pts)")
            total_score += 0.35
        elif full_name_correct >= 75:
            partial = round(0.35 * (full_name_correct / 80), 4)
            print(f"PASS (partial): Component 1 — Full Name formulas: {full_name_correct}/80 rows correct ({partial} pts)")
            total_score += partial
        elif full_name_correct >= 40:
            partial = round(0.35 * (full_name_correct / 80), 4)
            print(f"PASS (partial): Component 1 — Full Name formulas: {full_name_correct}/80 rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Full Name formulas: only {full_name_correct}/80 rows correct, expected =B{{n}}&\" \"&A{{n}}")
    except Exception as e:
        print(f"ERROR: Component 1 (Full Name) — {e}")

    # -------------------------------------------------------------------
    # Component 2: Formatted Phone column G (0.35 points)
    # Each row n should have formula: ="("&LEFT(C{n},3)&") "&MID(C{n},4,3)&"-"&RIGHT(C{n},4)
    # This was empty in the initial file — verifies task change.
    # -------------------------------------------------------------------
    try:
        phone_correct = 0
        # Pattern: ="("&LEFT(C{n},3)&") "&MID(C{n},4,3)&"-"&RIGHT(C{n},4)
        # We check for key structural elements: LEFT, MID, RIGHT, the parens and dashes
        phone_pattern = re.compile(
            r'^="?\("&LEFT\(C(\d+),3\)&"\) "&MID\(C\1,4,3\)&"-"&RIGHT\(C\1,4\)$',
            re.IGNORECASE
        )

        for row in range(2, 82):
            val = ws.cell(row=row, column=7).value  # Column G
            if val is not None and isinstance(val, str):
                val_stripped = val.strip()
                m = phone_pattern.match(val_stripped)
                if m and int(m.group(1)) == row:
                    phone_correct += 1
                else:
                    # Try alternate form without outer quotes on parens
                    expected = f'="("&LEFT(C{row},3)&") "&MID(C{row},4,3)&"-"&RIGHT(C{row},4)'
                    if val_stripped.upper().replace(' ', '') == expected.upper().replace(' ', ''):
                        phone_correct += 1

        if phone_correct == 80:
            print(f"PASS: Component 2 — Formatted Phone formulas: all 80/80 rows correct (0.35 pts)")
            total_score += 0.35
        elif phone_correct >= 75:
            partial = round(0.35 * (phone_correct / 80), 4)
            print(f"PASS (partial): Component 2 — Formatted Phone formulas: {phone_correct}/80 rows correct ({partial} pts)")
            total_score += partial
        elif phone_correct >= 40:
            partial = round(0.35 * (phone_correct / 80), 4)
            print(f"PASS (partial): Component 2 — Formatted Phone formulas: {phone_correct}/80 rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — Formatted Phone formulas: only {phone_correct}/80 rows correct")
            print(f"  Expected: =\"(\"&LEFT(C{{n}},3)&\") \"&MID(C{{n}},4,3)&\"-\"&RIGHT(C{{n}},4)")
    except Exception as e:
        print(f"ERROR: Component 2 (Formatted Phone) — {e}")

    # -------------------------------------------------------------------
    # Component 3: Grade Status column H (0.30 points)
    # Each row n should have IFS formula with all 4 grade thresholds:
    #   =IFS(E{n}>=3.5,"Honors",E{n}>=2.0,"Good Standing",E{n}>=1.0,"Academic Warning",TRUE,"Academic Probation")
    # This was empty in the initial file — verifies task change.
    # -------------------------------------------------------------------
    try:
        grade_correct = 0
        # Key check: formula must contain IFS, reference E column, and have all 4 status labels
        ifs_pattern = re.compile(r'=IFS\(', re.IGNORECASE)
        required_labels = ['"Honors"', '"Good Standing"', '"Academic Warning"', '"Academic Probation"']

        for row in range(2, 82):
            val = ws.cell(row=row, column=8).value  # Column H
            if val is not None and isinstance(val, str):
                val_stripped = val.strip()
                # Must be an IFS formula
                if not ifs_pattern.match(val_stripped):
                    continue
                # Must reference the E column for this row
                if f'E{row}' not in val_stripped:
                    continue
                # Must contain all 4 grade labels (case-insensitive check)
                val_upper = val_stripped.upper()
                all_labels_present = all(
                    label.upper() in val_upper for label in required_labels
                )
                if not all_labels_present:
                    continue
                # Must have the 3.5 threshold (Honors), 2.0 (Good Standing), 1.0 (Academic Warning)
                if '3.5' in val_stripped and '2.0' in val_stripped and '1.0' in val_stripped:
                    grade_correct += 1

        if grade_correct == 80:
            print(f"PASS: Component 3 — Grade Status IFS formulas: all 80/80 rows correct (0.30 pts)")
            total_score += 0.30
        elif grade_correct >= 75:
            partial = round(0.30 * (grade_correct / 80), 4)
            print(f"PASS (partial): Component 3 — Grade Status IFS formulas: {grade_correct}/80 rows correct ({partial} pts)")
            total_score += partial
        elif grade_correct >= 40:
            partial = round(0.30 * (grade_correct / 80), 4)
            print(f"PASS (partial): Component 3 — Grade Status IFS formulas: {grade_correct}/80 rows correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — Grade Status IFS formulas: only {grade_correct}/80 rows correct")
            print(f"  Expected IFS with thresholds 3.5/2.0/1.0 and labels Honors/Good Standing/Academic Warning/Academic Probation")
    except Exception as e:
        print(f"ERROR: Component 3 (Grade Status) — {e}")

    final_score = round(min(total_score, 1.0), 4)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
