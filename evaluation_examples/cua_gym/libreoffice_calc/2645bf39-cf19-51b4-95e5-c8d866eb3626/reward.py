"""
Reward Script: Verify pivot table counting survey responses per rating level
Task ID: calc_pivot_007
Domain: libreoffice_calc

Scoring Rubric:
  Component 1 (0.25): A new sheet exists (beyond the original 'Survey' sheet) containing pivot data
  Component 2 (0.30): Rating labels 1-5 are present as row fields in the pivot sheet
  Component 3 (0.30): Counts per rating match expected values (1=35, 2=52, 3=78, 4=85, 5=50)
  Component 4 (0.15): Grand Total row present with value 300
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_007'

# Expected counts from task context
EXPECTED_COUNTS = {1: 35, 2: 52, 3: 78, 4: 85, 5: 50}
EXPECTED_GRAND_TOTAL = 300


def find_pivot_sheet(wb):
    """Find the pivot/summary sheet (any sheet other than 'Survey')."""
    for sn in wb.sheetnames:
        if sn.lower() != 'survey':
            return wb[sn]
    return None


def find_data_in_sheet(ws):
    """
    Search the pivot sheet for rating-count pairs and a grand total.
    Returns (rating_counts_dict, grand_total_value).
    Scans all rows looking for rating values 1-5 paired with numeric counts,
    and a row with 'Grand Total' paired with a numeric value.
    """
    rating_counts = {}
    grand_total = None

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        cells = [c.value for c in row]
        # Look for grand total row
        for i, val in enumerate(cells):
            if val is not None and str(val).lower().replace(' ', '') == 'grandtotal':
                # Find the numeric value in this row
                for j, v2 in enumerate(cells):
                    if j != i and isinstance(v2, (int, float)):
                        grand_total = int(v2)
                        break
                continue

        # Look for rating-count pairs: an integer 1-5 paired with a count
        numeric_cells = [(i, cells[i]) for i in range(len(cells))
                         if isinstance(cells[i], (int, float))]
        for idx, val in numeric_cells:
            int_val = int(val)
            if int_val in range(1, 6) and int_val not in rating_counts:
                # Find the paired count value (another numeric in the same row)
                for idx2, val2 in numeric_cells:
                    if idx2 != idx and isinstance(val2, (int, float)):
                        rating_counts[int_val] = int(val2)
                        break

    return rating_counts, grand_total


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: A pivot/summary sheet exists beyond 'Survey' (0.25 points)
    # This is the core structural change - initial has only 'Survey'
    try:
        pivot_ws = find_pivot_sheet(wb)
        if pivot_ws is not None:
            print(f"PASS: Component 1 — Found pivot sheet '{pivot_ws.title}' (0.25 pts)")
            total_score += 0.25
        else:
            print("FAIL: Component 1 — No sheet other than 'Survey' found")
            # If no pivot sheet, nothing else to check
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {total_score}")
            return total_score
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {total_score}")
        return total_score

    # Component 2: Rating labels 1-5 present as row fields (0.30 points)
    try:
        rating_counts, grand_total = find_data_in_sheet(pivot_ws)
        found_ratings = set(rating_counts.keys())
        expected_ratings = {1, 2, 3, 4, 5}
        if found_ratings == expected_ratings:
            print(f"PASS: Component 2 — All ratings 1-5 found as row fields (0.30 pts)")
            total_score += 0.30
        elif len(found_ratings) > 0:
            partial = 0.30 * len(found_ratings.intersection(expected_ratings)) / 5
            if partial > 0:
                print(f"PARTIAL: Component 2 — Found ratings {sorted(found_ratings)}, "
                      f"expected {{1,2,3,4,5}} ({partial:.2f} pts)")
                total_score += partial
        else:
            print("FAIL: Component 2 — No rating labels found in pivot sheet")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Counts per rating match expected values (0.30 points)
    # Expected: 1=35, 2=52, 3=78, 4=85, 5=50
    try:
        correct_count = 0
        for rating in range(1, 6):
            actual = rating_counts.get(rating)
            expected = EXPECTED_COUNTS[rating]
            if actual == expected:
                correct_count += 1
                print(f"  Rating {rating}: count={actual} == {expected} OK")
            else:
                print(f"  Rating {rating}: count={actual} != expected {expected}")

        if correct_count == 5:
            print(f"PASS: Component 3 — All 5 rating counts correct (0.30 pts)")
            total_score += 0.30
        elif correct_count > 0:
            partial = 0.30 * correct_count / 5
            if partial > 0:
                print(f"PARTIAL: Component 3 — {correct_count}/5 counts correct ({partial:.2f} pts)")
                total_score += partial
        else:
            print("FAIL: Component 3 — No rating counts match expected values")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # Component 4: Grand Total row present with value 300 (0.15 points)
    try:
        if grand_total == EXPECTED_GRAND_TOTAL:
            print(f"PASS: Component 4 — Grand Total = {grand_total} (0.15 pts)")
            total_score += 0.15
        else:
            print(f"FAIL: Component 4 — Grand Total = {grand_total}, expected {EXPECTED_GRAND_TOTAL}")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point — persist app state then verify
def persist_app_state(domain):
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
