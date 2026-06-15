"""
Initial Setup: Survey results spreadsheet for pivot table task
Task ID: calc_pivot_007
Domain: libreoffice_calc
"""

import os
import random
import shlex
import subprocess
import time

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_pivot_007'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'

def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)

def create_initial():
    random.seed(42)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Survey'

    # Headers
    headers = ['ResponseID', 'Respondent', 'Question', 'Rating', 'Comment']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic names pool
    first_names = [
        'Sarah', 'Marcus', 'Elena', 'James', 'Priya', 'Carlos', 'Megan',
        'David', 'Aisha', 'Ryan', 'Yuki', 'Thomas', 'Fatima', 'Kevin',
        'Sophia', 'Andre', 'Lina', 'Robert', 'Chen', 'Jessica',
        'Miguel', 'Hannah', 'Omar', 'Rachel', 'Dmitri', 'Angela',
        'Hiroshi', 'Laura', 'Samuel', 'Natasha', 'Patrick', 'Maria',
        'Jonathan', 'Deepa', 'Brian', 'Olivia', 'Ibrahim', 'Kate',
        'Vincent', 'Zara', 'Nathan', 'Emma', 'Luis', 'Amara',
        'Daniel', 'Chloe', 'Wei', 'Grace', 'Alexander', 'Nadia'
    ]
    last_names = [
        'Chen', 'Johnson', 'Petrova', 'Williams', 'Sharma', 'Garcia',
        'O\'Brien', 'Kim', 'Hassan', 'Taylor', 'Tanaka', 'Brown',
        'Ali', 'Lee', 'Rossi', 'Dubois', 'Muller', 'Singh', 'Wang',
        'Martinez', 'Nakamura', 'Anderson', 'Okafor', 'Wilson',
        'Kozlov', 'Davis', 'Yamamoto', 'Moore', 'Abe', 'Fernandez',
        'Thompson', 'Gupta', 'Clark', 'Rivera', 'Johansson', 'Patel',
        'Bailey', 'Ng', 'Fischer', 'Moreau'
    ]

    questions = ['Q1', 'Q2', 'Q3', 'Q4', 'Q5']

    # Comments pool by rating
    comments_by_rating = {
        1: [
            'Very disappointed with the service',
            'Needs major improvements',
            'Not meeting expectations at all',
            'Would not recommend to others',
            'Extremely unsatisfied with the experience',
            'Below acceptable standards',
            'Poor quality throughout',
        ],
        2: [
            'Below average experience',
            'Some aspects need work',
            'Not quite what I expected',
            'Room for significant improvement',
            'Somewhat disappointing overall',
            'Could be much better',
            'Fell short of expectations',
        ],
        3: [
            'Average experience overall',
            'Meets basic expectations',
            'Neither good nor bad',
            'Acceptable but unremarkable',
            'Middle of the road quality',
            'Some good and some bad aspects',
            'Fair performance overall',
        ],
        4: [
            'Good experience overall',
            'Above average quality',
            'Would recommend with minor reservations',
            'Mostly satisfied with the results',
            'Solid performance across the board',
            'Impressed with most aspects',
            'Very good but not perfect',
        ],
        5: [
            'Excellent experience all around',
            'Exceeded all expectations',
            'Highly recommend to everyone',
            'Outstanding quality and service',
            'Absolutely wonderful experience',
            'Best in class performance',
            'Could not be more satisfied',
        ],
    }

    # Build rating distribution: 1=35, 2=52, 3=78, 4=85, 5=50
    ratings = (
        [1] * 35 +
        [2] * 52 +
        [3] * 78 +
        [4] * 85 +
        [5] * 50
    )
    random.shuffle(ratings)

    # Generate 300 rows
    for i in range(300):
        row = i + 2  # data starts at row 2
        response_id = i + 1
        first = random.choice(first_names)
        last = random.choice(last_names)
        respondent = f'{first} {last}'
        question = random.choice(questions)
        rating = ratings[i]
        comment = random.choice(comments_by_rating[rating])

        ws.cell(row=row, column=1, value=response_id)
        ws.cell(row=row, column=2, value=respondent)
        ws.cell(row=row, column=3, value=question)
        ws.cell(row=row, column=4, value=rating)
        ws.cell(row=row, column=5, value=comment)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 40

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open LibreOffice Calc with the file
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')

create_initial()
