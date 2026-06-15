"""
Reward Script: Filter customer database to show only gmail.com email addresses
Task ID: calc_dop_filter_contains_012
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5): AutoFilter has a custom filter on Email column (col C / index 2)
                      filtering for values containing 'gmail.com'
  Component 2 (0.5): Exactly 67 rows are visible and all visible rows contain gmail.com;
                      133 non-gmail rows are hidden
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — reward scripts run on the VM
TASK_ID = 'calc_dop_filter_contains_012'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Customers sheet must exist
    if 'Customers' not in wb.sheetnames:
        print("CRITICAL: 'Customers' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Customers']

    # Component 1: AutoFilter has a custom filter on Email column (index 2) for gmail.com (0.5 points)
    # The task requires applying a filter on column C (Email) containing 'gmail.com'.
    # In openpyxl, filterColumn uses 0-based colId; Email is column C = index 2.
    # A standard "Contains gmail.com" filter appears as a CustomFilter with val='*gmail.com*'.
    try:
        filter_applied = False
        filter_detail = None

        if ws.auto_filter.ref and ws.auto_filter.filterColumn:
            for fc in ws.auto_filter.filterColumn:
                if fc.colId == 2:
                    # Check for custom filter containing gmail.com
                    if fc.customFilters is not None:
                        for cf in fc.customFilters.customFilter:
                            val = str(cf.val).lower() if cf.val else ''
                            if 'gmail.com' in val:
                                filter_applied = True
                                filter_detail = f"colId={fc.colId}, val={cf.val}, operator={cf.operator}"
                                break
                    # Also check plain filters (discrete value list)
                    if not filter_applied and fc.filters is not None:
                        filter_vals = [str(f.val).lower() for f in fc.filters.filter] if fc.filters.filter else []
                        if any('gmail.com' in v for v in filter_vals):
                            filter_applied = True
                            filter_detail = f"colId={fc.colId}, filter values={filter_vals[:3]}"

        if filter_applied:
            print(f"PASS: Component 1 — AutoFilter on Email column contains gmail.com filter ({filter_detail}) (0.5 pts)")
            total_score += 0.5
        else:
            # Show what filters exist for debugging
            if ws.auto_filter.filterColumn:
                fc_info = [(fc.colId, str(fc.customFilters)[:80] if fc.customFilters else str(fc.filters)[:80])
                           for fc in ws.auto_filter.filterColumn]
                print(f"FAIL: Component 1 — No gmail.com filter on Email column (colId=2). Found: {fc_info}")
            else:
                print(f"FAIL: Component 1 — No filterColumn entries in AutoFilter. auto_filter.ref={ws.auto_filter.ref}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Row visibility — exactly 67 gmail.com rows visible, 133 non-gmail rows hidden (0.5 points)
    # The task says ~67 rows should be visible (those with gmail.com emails).
    # We verify: (a) 133 rows are hidden and (b) all visible rows have gmail.com email addresses.
    try:
        hidden_rows = []
        visible_data_rows = []

        for row_idx in range(2, 202):  # rows 2-201 are data rows
            rd = ws.row_dimensions.get(row_idx)
            is_hidden = rd is not None and rd.hidden
            if is_hidden:
                hidden_rows.append(row_idx)
            else:
                visible_data_rows.append(row_idx)

        # Check that visible rows all contain gmail.com
        visible_non_gmail = []
        for row_idx in visible_data_rows:
            email = ws.cell(row=row_idx, column=3).value
            if email and 'gmail.com' not in str(email).lower():
                visible_non_gmail.append((row_idx, email))

        hidden_count = len(hidden_rows)
        visible_count = len(visible_data_rows)

        # Expect 133 hidden, 67 visible, and 0 visible non-gmail rows
        if hidden_count == 133 and visible_count == 67 and len(visible_non_gmail) == 0:
            print(f"PASS: Component 2 — {hidden_count} rows hidden, {visible_count} visible, "
                  f"all visible rows have gmail.com emails (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 2 — hidden={hidden_count} (expected 133), "
                  f"visible={visible_count} (expected 67), "
                  f"visible non-gmail={len(visible_non_gmail)} (expected 0)")
            if visible_non_gmail:
                print(f"  Sample visible non-gmail: {visible_non_gmail[:3]}")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
