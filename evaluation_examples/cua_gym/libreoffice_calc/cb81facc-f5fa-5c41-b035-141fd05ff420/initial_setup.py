"""
Initial Setup: Customer database with AutoFilter enabled, no filter applied
Task ID: calc_dop_filter_contains_012
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.utils import get_column_letter

WORKDIR = '/home/user'
TASK_ID = 'calc_dop_filter_contains_012'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Customers'

    # Headers
    headers = ['Customer ID', 'Name', 'Email', 'Phone', 'City', 'Join Date', 'Total Purchases']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic customer data: ~67 gmail.com out of 200
    # Mix of domains: gmail.com, yahoo.com, outlook.com, hotmail.com, company.com, business.net
    first_names = [
        'Emma', 'Liam', 'Olivia', 'Noah', 'Ava', 'William', 'Sophia', 'James',
        'Isabella', 'Oliver', 'Mia', 'Benjamin', 'Charlotte', 'Elijah', 'Amelia',
        'Lucas', 'Harper', 'Mason', 'Evelyn', 'Logan', 'Abigail', 'Ethan', 'Emily',
        'Aiden', 'Elizabeth', 'Jackson', 'Sofia', 'Sebastian', 'Avery', 'Mateo',
        'Ella', 'Jack', 'Scarlett', 'Owen', 'Grace', 'Theodore', 'Chloe', 'Amir',
        'Victoria', 'Wyatt', 'Riley', 'Charles', 'Aria', 'Julian', 'Lily', 'Levi',
        'Aurora', 'Isaac', 'Zoey', 'Gabriel', 'Nora', 'Anthony', 'Hazel', 'Dylan',
        'Luna', 'Lincoln', 'Camila', 'Jayden', 'Penelope', 'Christopher', 'Layla',
        'Joshua', 'Hannah', 'Andrew', 'Lillian', 'Leo', 'Addison', 'John', 'Aubrey',
        'Ryan', 'Eleanor', 'Nathan', 'Ellie', 'Carter', 'Stella', 'Caleb', 'Natalia',
        'Luke', 'Zoe', 'Isaiah', 'Leah', 'Henry', 'Hailee', 'Eli', 'Violet',
        'Aaron', 'Clara', 'Daniel', 'Paisley', 'Adam', 'Everly', 'Landon', 'Autumn',
        'Robert', 'Savannah', 'Thomas', 'Brooklyn'
    ]

    last_names = [
        'Wilson', 'Martinez', 'Anderson', 'Taylor', 'Thomas', 'Jackson', 'White',
        'Harris', 'Thompson', 'Garcia', 'Moore', 'Jones', 'Davis', 'Miller',
        'Rodriguez', 'Smith', 'Brown', 'Johnson', 'Williams', 'Lee', 'Walker',
        'Hall', 'Allen', 'Young', 'King', 'Wright', 'Scott', 'Green', 'Baker',
        'Adams', 'Nelson', 'Carter', 'Mitchell', 'Perez', 'Roberts', 'Turner',
        'Phillips', 'Campbell', 'Parker', 'Evans', 'Collins', 'Stewart', 'Sanchez',
        'Morris', 'Rogers', 'Reed', 'Cook', 'Morgan', 'Bell', 'Murphy', 'Bailey',
        'Rivera', 'Cooper', 'Richardson', 'Cox', 'Howard', 'Ward', 'Torres',
        'Peterson', 'Gray', 'Ramirez', 'James', 'Watson', 'Brooks', 'Kelly',
        'Sanders', 'Price', 'Bennett', 'Wood', 'Barnes', 'Ross', 'Henderson',
        'Coleman', 'Jenkins', 'Perry', 'Powell', 'Long', 'Patterson', 'Hughes',
        'Flores', 'Washington', 'Butler', 'Simmons', 'Foster', 'Gonzales', 'Bryant',
        'Alexander', 'Russell', 'Griffin', 'Diaz', 'Hayes', 'Myers', 'Ford',
        'Hamilton', 'Graham', 'Sullivan', 'Wallace', 'West', 'Cole', 'Jordan'
    ]

    cities = [
        'Austin', 'Denver', 'Seattle', 'Portland', 'Nashville', 'Boston', 'Chicago',
        'Atlanta', 'Miami', 'Phoenix', 'Dallas', 'Houston', 'San Diego', 'Los Angeles',
        'New York', 'Philadelphia', 'San Francisco', 'Las Vegas', 'Minneapolis',
        'Detroit', 'Charlotte', 'Baltimore', 'Indianapolis', 'Columbus', 'Memphis',
        'Louisville', 'Oklahoma City', 'Tucson', 'Sacramento', 'Fresno'
    ]

    # Domain assignment: ~67 gmail, ~50 yahoo, ~40 outlook, ~25 hotmail, ~18 company/business
    # Total 200 rows
    # Indices 0-66: gmail.com (67)
    # Indices 67-116: yahoo.com (50)
    # Indices 117-156: outlook.com (40)
    # Indices 157-181: hotmail.com (25)
    # Indices 182-191: company.com (10)
    # Indices 192-199: business.net (8)

    def get_domain(idx):
        if idx < 67:
            return 'gmail.com'
        elif idx < 117:
            return 'yahoo.com'
        elif idx < 157:
            return 'outlook.com'
        elif idx < 182:
            return 'hotmail.com'
        elif idx < 192:
            return 'company.com'
        else:
            return 'business.net'

    # Generate dates spanning 2020-2024
    import datetime
    base_date = datetime.date(2020, 1, 1)
    total_days = (datetime.date(2025, 1, 1) - base_date).days  # ~1826 days

    import random
    random.seed(42)

    rows_data = []
    used_names = set()
    for i in range(200):
        # Pick unique name
        attempts = 0
        while True:
            fn = first_names[i % len(first_names)]
            ln = last_names[(i * 3 + attempts) % len(last_names)]
            full_name = f'{fn} {ln}'
            if full_name not in used_names:
                used_names.add(full_name)
                break
            attempts += 1

        customer_id = f'C{i+1:03d}'
        email_user = f'{fn.lower()}.{ln.lower()}{i if i >= 50 else ""}'
        # Clean email user (remove spaces, numbers for first 50)
        if i < 50:
            email_user = f'{fn.lower()}.{ln.lower()}'
        else:
            email_user = f'{fn.lower()}.{ln.lower()}{i}'
        domain = get_domain(i)
        email = f'{email_user}@{domain}'

        # Phone
        area = random.randint(200, 999)
        phone = f'{area}-{random.randint(1000,9999):04d}-{random.randint(1000,9999):04d}'

        city = cities[i % len(cities)]

        # Join date
        day_offset = int((i / 200) * total_days) + random.randint(0, 30)
        join_date = (base_date + datetime.timedelta(days=day_offset)).strftime('%Y-%m-%d')

        # Total purchases
        total_purchases = round(random.uniform(50.0, 5000.0), 2)

        rows_data.append([customer_id, full_name, email, phone, city, join_date, total_purchases])

    # Write data rows
    for r, row_data in enumerate(rows_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Enable AutoFilter on full range (headers + data = rows 1-201, columns A-G)
    ws.auto_filter.ref = 'A1:G201'

    # Set column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 18

    # Freeze the header row
    ws.freeze_panes = 'A2'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Rows: 200 data rows (rows 2-201)')
    print(f'  Columns: Customer ID, Name, Email, Phone, City, Join Date, Total Purchases')
    print(f'  AutoFilter: enabled on A1:G201, no rows hidden')


create_initial()
