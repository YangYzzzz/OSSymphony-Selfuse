"""
Reward Script: Build a comprehensive employee onboarding checklist with progress tracking
Task ID: calc_gpm_080
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20) - Title row A1:H1 merged with styling (bold, 14pt, centered, white on indigo)
  Component 2 (0.20) - Phase cells merged in column A (4 phase groups)
  Component 3 (0.20) - G5:G20 contain =D<row>-TODAY() formulas
  Component 4 (0.15) - Data validations: status dropdown (E5:E20) and completed dropdown (F5:F20)
  Component 5 (0.10) - Conditional formatting on E5:E20 and G5:G20
  Component 6 (0.10) - Row 22 completion progress label + formula
  Component 7 (0.05) - Row 4 headers bold with indigo fill
"""

import os
import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_080'


def persist_app_state(domain):
    """Try to save any unsaved LibreOffice state via Ctrl+S."""
    os.environ["DISPLAY"] = ":0"
    try:
        import pyautogui
        pyautogui.hotkey("ctrl", "s")
        import time
        time.sleep(0.8)
        print("PERSIST: ctrl+s sent for libreoffice_calc")
    except Exception as e:
        print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    if 'Onboard' not in wb.sheetnames:
        print("FAIL: Sheet 'Onboard' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Onboard']
    merged_ranges = [str(r) for r in ws.merged_cells.ranges]

    # Component 1: Title row A1:H1 merged with styling (0.20 points)
    # Checks: A1:H1 merged, A1 bold, A1 size ~14, A1 centered, A1 white font on dark fill
    try:
        comp1 = 0.0
        # Check merge
        has_title_merge = any('A1' in r and 'H1' in r for r in merged_ranges)
        if has_title_merge:
            comp1 += 0.08
            print("PASS: Component 1a - A1:H1 is merged (0.08 pts)")
        else:
            print(f"FAIL: Component 1a - A1:H1 not merged. Ranges: {merged_ranges}")

        cell_a1 = ws['A1']
        # Check bold and size
        if cell_a1.font.bold and cell_a1.font.size is not None and cell_a1.font.size >= 13:
            comp1 += 0.04
            print(f"PASS: Component 1b - A1 bold={cell_a1.font.bold}, size={cell_a1.font.size} (0.04 pts)")
        else:
            print(f"FAIL: Component 1b - A1 bold={cell_a1.font.bold}, size={cell_a1.font.size}")

        # Check centered alignment
        if cell_a1.alignment.horizontal == 'center':
            comp1 += 0.04
            print("PASS: Component 1c - A1 horizontally centered (0.04 pts)")
        else:
            print(f"FAIL: Component 1c - A1 alignment.horizontal={cell_a1.alignment.horizontal}")

        # Check fill (dark indigo-ish background)
        try:
            fill_rgb = cell_a1.fill.fgColor.rgb
            if fill_rgb and cell_a1.fill.fill_type == 'solid':
                comp1 += 0.04
                print(f"PASS: Component 1d - A1 has solid fill ({fill_rgb}) (0.04 pts)")
            else:
                print(f"FAIL: Component 1d - A1 fill type={cell_a1.fill.fill_type}, rgb={fill_rgb}")
        except Exception:
            print("FAIL: Component 1d - Could not read A1 fill")

        if comp1 > 0:
            total_score += comp1
            print(f"Component 1 subtotal: {comp1}/0.20")
    except Exception as e:
        print(f"ERROR: Component 1 - {e}")

    # Component 2: Phase cells merged in column A (0.20 points)
    # Expected merges: A5:A8 (Pre-boarding), A9:A12 (Day 1), A13:A16 (Week 1), A17:A20 (Month 1)
    try:
        expected_phase_merges = ['A5:A8', 'A9:A12', 'A13:A16', 'A17:A20']
        phase_merge_count = 0
        for em in expected_phase_merges:
            if any(em in r for r in merged_ranges):
                phase_merge_count += 1

        if phase_merge_count == 4:
            total_score += 0.20
            print(f"PASS: Component 2 - All 4 phase merges found ({phase_merge_count}/4) (0.20 pts)")
        elif phase_merge_count > 0:
            partial = 0.20 * (phase_merge_count / 4)
            total_score += partial
            print(f"PARTIAL: Component 2 - {phase_merge_count}/4 phase merges found ({partial:.2f} pts)")
        else:
            print(f"FAIL: Component 2 - No phase merges found in {merged_ranges}")
    except Exception as e:
        print(f"ERROR: Component 2 - {e}")

    # Component 3: G5:G20 contain formulas like =D<row>-TODAY() (0.20 points)
    try:
        formula_count = 0
        for row_num in range(5, 21):
            cell_val = ws.cell(row=row_num, column=7).value
            if cell_val is not None and isinstance(cell_val, str) and 'TODAY()' in cell_val.upper():
                formula_count += 1

        if formula_count >= 14:
            total_score += 0.20
            print(f"PASS: Component 3 - {formula_count}/16 G-column formulas contain TODAY() (0.20 pts)")
        elif formula_count > 0:
            partial = 0.20 * (formula_count / 16)
            total_score += partial
            print(f"PARTIAL: Component 3 - {formula_count}/16 G-column formulas ({partial:.2f} pts)")
        else:
            print(f"FAIL: Component 3 - No TODAY() formulas found in G5:G20")
    except Exception as e:
        print(f"ERROR: Component 3 - {e}")

    # Component 4: Data validations for Status and Completed dropdowns (0.15 points)
    try:
        dv_list = ws.data_validations.dataValidation if ws.data_validations else []
        has_status_dv = False
        has_completed_dv = False

        for dv in dv_list:
            formula_str = str(dv.formula1) if dv.formula1 else ''
            sqref_str = str(dv.sqref) if dv.sqref else ''
            # Check for status dropdown (Not Started, In Progress, Complete, Blocked)
            if dv.type == 'list' and ('Not Started' in formula_str or 'Complete' in formula_str or 'Blocked' in formula_str):
                has_status_dv = True
            # Check for completed dropdown (Yes/No)
            if dv.type == 'list' and ('Yes' in formula_str and 'No' in formula_str):
                has_completed_dv = True

        comp4 = 0.0
        if has_status_dv:
            comp4 += 0.10
            print("PASS: Component 4a - Status dropdown validation found (0.10 pts)")
        else:
            print("FAIL: Component 4a - No status dropdown found")

        if has_completed_dv:
            comp4 += 0.05
            print("PASS: Component 4b - Yes/No completed dropdown found (0.05 pts)")
        else:
            print("FAIL: Component 4b - No Yes/No dropdown found")

        total_score += comp4
    except Exception as e:
        print(f"ERROR: Component 4 - {e}")

    # Component 5: Conditional formatting on E and G columns (0.10 points)
    try:
        cf_ranges = []
        for cf in ws.conditional_formatting:
            cf_ranges.append(str(cf))

        has_e_cf = any('E' in r for r in cf_ranges)
        has_g_cf = any('G' in r for r in cf_ranges)

        comp5 = 0.0
        if has_e_cf:
            comp5 += 0.05
            print(f"PASS: Component 5a - Conditional formatting on E column (0.05 pts)")
        else:
            print(f"FAIL: Component 5a - No conditional formatting on E column. Ranges: {cf_ranges}")

        if has_g_cf:
            comp5 += 0.05
            print(f"PASS: Component 5b - Conditional formatting on G column (0.05 pts)")
        else:
            print(f"FAIL: Component 5b - No conditional formatting on G column. Ranges: {cf_ranges}")

        total_score += comp5
    except Exception as e:
        print(f"ERROR: Component 5 - {e}")

    # Component 6: Row 22 completion progress label + formula (0.10 points)
    try:
        comp6 = 0.0
        a22_val = ws['A22'].value
        b22_val = ws['B22'].value

        if a22_val and 'completion' in str(a22_val).lower() and 'progress' in str(a22_val).lower():
            comp6 += 0.04
            print(f"PASS: Component 6a - A22 has completion progress label: '{a22_val}' (0.04 pts)")
        else:
            print(f"FAIL: Component 6a - A22 value: {repr(a22_val)}")

        if b22_val and isinstance(b22_val, str) and 'COUNTIF' in b22_val.upper():
            comp6 += 0.04
            print(f"PASS: Component 6b - B22 has COUNTIF formula: '{b22_val}' (0.04 pts)")
        else:
            print(f"FAIL: Component 6b - B22 value: {repr(b22_val)}")

        # Check A22 is bold and 12pt
        if ws['A22'].font.bold and ws['A22'].font.size is not None and ws['A22'].font.size >= 12:
            comp6 += 0.02
            print(f"PASS: Component 6c - A22 bold 12pt (0.02 pts)")
        else:
            print(f"FAIL: Component 6c - A22 bold={ws['A22'].font.bold}, size={ws['A22'].font.size}")

        total_score += comp6
    except Exception as e:
        print(f"ERROR: Component 6 - {e}")

    # Component 7: Row 4 headers bold with indigo fill (0.05 points)
    try:
        comp7 = 0.0
        a4 = ws['A4']
        if a4.font.bold:
            comp7 += 0.025
            print("PASS: Component 7a - Row 4 headers bold (0.025 pts)")
        else:
            print(f"FAIL: Component 7a - A4 bold={a4.font.bold}")

        try:
            fill_rgb = a4.fill.fgColor.rgb
            if fill_rgb and a4.fill.fill_type == 'solid':
                comp7 += 0.025
                print(f"PASS: Component 7b - Row 4 has solid fill ({fill_rgb}) (0.025 pts)")
            else:
                print(f"FAIL: Component 7b - A4 fill_type={a4.fill.fill_type}")
        except Exception:
            print("FAIL: Component 7b - Could not read A4 fill")

        total_score += comp7
    except Exception as e:
        print(f"ERROR: Component 7 - {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
