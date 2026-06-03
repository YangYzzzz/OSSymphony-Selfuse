"""
Initial Setup: Employee onboarding checklist - raw data only
Task ID: calc_gpm_080
Domain: libreoffice_calc

The initial file contains the raw task data (employee info, task list)
but WITHOUT the checklist formatting features: no merged cells, no title
formatting, no data validation dropdowns, no conditional formatting,
no formulas, no progress tracking row. The agent must build all of that.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_080'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Onboard'

    # --- Row 1: Plain title text (NO merge, NO formatting) ---
    ws['A1'] = 'New Employee Onboarding Checklist'

    # --- Row 2: Employee info (plain text, NO italic formatting) ---
    ws['A2'] = 'Employee: Sarah Chen'
    ws['B2'] = 'Start Date: April 7, 2026'
    ws['D2'] = 'Department: Engineering'
    ws['F2'] = 'Manager: Tom Rodriguez'

    # --- Row 3: blank ---

    # --- Row 4: Headers (plain text, NO formatting) ---
    headers = ['Phase', 'Task', 'Owner', 'Due Date', 'Status',
               'Completed', 'Days Until Due', 'Notes']
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)

    # --- Task Data (rows 5-20) ---
    # Phase, Task, Owner, Due Date, Status, Completed, Days Until Due, Notes
    tasks = [
        # Pre-boarding (rows 5-8)
        ['Pre-boarding', 'Set up workstation and equipment', 'IT Support', 'Apr 01', 'Not Started', 'No', '', 'Order laptop and peripherals'],
        ['Pre-boarding', 'Assign desk and office space', 'Facilities', 'Apr 02', 'Not Started', 'No', '', 'Building B, Floor 3'],
        ['Pre-boarding', 'Send welcome email with first day info', 'HR Team', 'Apr 03', 'Not Started', 'No', '', 'Include parking and dress code'],
        ['Pre-boarding', 'Request employee badge and access card', 'Security', 'Apr 04', 'Not Started', 'No', '', 'Photo needed from HR'],

        # Day 1 (rows 9-12)
        ['Day 1', 'Company orientation session', 'HR Team', 'Apr 07', 'Not Started', 'No', '', '9:00 AM in Conference Room A'],
        ['Day 1', 'Team introduction and office tour', 'Tom Rodriguez', 'Apr 07', 'Not Started', 'No', '', 'Meet engineering team'],
        ['Day 1', 'Set up email and software accounts', 'IT Support', 'Apr 07', 'Not Started', 'No', '', 'GitHub, Slack, Jira access'],
        ['Day 1', 'Complete HR paperwork and benefits forms', 'HR Team', 'Apr 07', 'Not Started', 'No', '', 'I-9, W-4, direct deposit'],

        # Week 1 (rows 13-16)
        ['Week 1', 'Complete required training modules', 'Training Dept', 'Apr 11', 'Not Started', 'No', '', 'Safety and compliance courses'],
        ['Week 1', 'Assign mentor from engineering team', 'Tom Rodriguez', 'Apr 09', 'Not Started', 'No', '', 'Pair with senior engineer'],
        ['Week 1', 'Grant access to development tools', 'IT Support', 'Apr 08', 'Not Started', 'No', '', 'CI/CD pipeline, staging env'],
        ['Week 1', 'First project briefing and goals', 'Tom Rodriguez', 'Apr 11', 'Not Started', 'No', '', 'Q2 backend migration project'],

        # Month 1 (rows 17-20)
        ['Month 1', '30-day performance review', 'Tom Rodriguez', 'May 07', 'Not Started', 'No', '', 'Review goals and feedback'],
        ['Month 1', 'Complete security awareness training', 'Security', 'May 02', 'Not Started', 'No', '', 'Annual security certification'],
        ['Month 1', 'Benefits enrollment deadline', 'HR Team', 'Apr 30', 'Not Started', 'No', '', 'Health, dental, 401k options'],
        ['Month 1', 'Team integration and social events', 'Tom Rodriguez', 'May 05', 'Not Started', 'No', '', 'Lunch with cross-functional teams'],
    ]

    for r, row_data in enumerate(tasks, 5):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # --- Set reasonable column widths so data is visible ---
    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 16
    ws.column_dimensions['H'].width = 35

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
