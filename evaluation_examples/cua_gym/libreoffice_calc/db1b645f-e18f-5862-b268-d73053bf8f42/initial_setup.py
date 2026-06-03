"""
Initial Setup: INDEX-MATCH lookup in a pivot-style table
Task ID: calc_lf_048
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_048'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'StoreSales'

    # Row 1: Headers (A1 empty, B1-D1 store names)
    ws.cell(row=1, column=2, value='Store NYC')
    ws.cell(row=1, column=3, value='Store LA')
    ws.cell(row=1, column=4, value='Store CHI')

    # Row 2: Shoes
    ws.cell(row=2, column=1, value='Shoes')
    ws.cell(row=2, column=2, value=5200)
    ws.cell(row=2, column=3, value=6100)
    ws.cell(row=2, column=4, value=4300)

    # Row 3: Shirts
    ws.cell(row=3, column=1, value='Shirts')
    ws.cell(row=3, column=2, value=3800)
    ws.cell(row=3, column=3, value=4500)
    ws.cell(row=3, column=4, value=3200)

    # Row 4: Pants
    ws.cell(row=4, column=1, value='Pants')
    ws.cell(row=4, column=2, value=4100)
    ws.cell(row=4, column=3, value=4900)
    ws.cell(row=4, column=4, value=3700)

    # Lookup section: F1:H1 headers, F2:G2 lookup values, H2 empty (task target)
    ws.cell(row=1, column=6, value='Product')
    ws.cell(row=2, column=6, value='Shirts')
    ws.cell(row=1, column=7, value='Store')
    ws.cell(row=2, column=7, value='Store LA')
    ws.cell(row=1, column=8, value='Sales')
    # H2 intentionally left empty - the agent must enter the formula

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
