"""
Reward Script: Two-way lookup using INDEX/MATCH in a pivot-style table
Task ID: calc_lf_048
Domain: libreoffice_calc
Scoring:
  Component 1 (0.4 pts): H2 contains a formula (not empty/literal)
  Component 2 (0.3 pts): Formula uses INDEX+MATCH two-way lookup pattern
  Component 3 (0.3 pts): Formula references correct ranges and yields 4500
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_048'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: StoreSales sheet must exist
    if 'StoreSales' not in wb.sheetnames:
        print("CRITICAL: 'StoreSales' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['StoreSales']

    # Component 1: H2 contains a formula (0.4 points)
    # This is the core task-introduced change: H2 goes from empty to having a formula.
    try:
        h2_value = ws['H2'].value
        if h2_value is not None and isinstance(h2_value, str) and h2_value.startswith('='):
            print(f"PASS: Component 1 — H2 contains a formula: {h2_value} (0.4 pts)")
            total_score += 0.4
        else:
            print(f"FAIL: Component 1 — H2 does not contain a formula. Found: {h2_value!r}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: Formula uses INDEX+MATCH two-way lookup pattern (0.3 points)
    # The task specifically asks for a two-way lookup using INDEX/MATCH.
    try:
        h2_value = ws['H2'].value
        if h2_value is not None and isinstance(h2_value, str):
            formula_upper = h2_value.upper().replace(' ', '')
            has_index = 'INDEX(' in formula_upper
            # Count MATCH occurrences — two-way lookup needs at least 2 MATCH calls
            match_count = formula_upper.count('MATCH(')
            if has_index and match_count >= 2:
                print(f"PASS: Component 2 — Formula uses INDEX with {match_count} MATCH calls (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 2 — Expected INDEX+2xMATCH pattern. INDEX={has_index}, MATCH count={match_count}")
        else:
            print(f"FAIL: Component 2 — H2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Formula references correct ranges and would produce 4500 (0.3 points)
    # Ground truth: =INDEX(B2:D4,MATCH(F2,A2:A4,0),MATCH(G2,B1:D1,0))
    # F2='Shirts', G2='Store LA' → row match on A2:A4 finds 'Shirts' at row 2 (of range),
    # col match on B1:D1 finds 'Store LA' at col 2 (of range) → B2:D4[2,2] = C3 = 4500
    # We verify the formula references the data range and lookup parameters correctly.
    try:
        h2_value = ws['H2'].value
        if h2_value is not None and isinstance(h2_value, str):
            formula_norm = h2_value.upper().replace(' ', '')

            # Check that INDEX references the data array (B2:D4 or equivalent)
            # The data range should cover B2:D4 (the numeric values)
            has_data_range = bool(re.search(r'INDEX\(B2:D4', formula_norm))

            # Check MATCH references: one for rows (A2:A4) and one for columns (B1:D1)
            has_row_match = bool(re.search(r'MATCH\(F2,A2:A4', formula_norm))
            has_col_match = bool(re.search(r'MATCH\(G2,B1:D1', formula_norm))

            # Also verify by manual computation that the result would be 4500
            # F2='Shirts' is at index 2 in A2:A4 (A2=Shoes, A3=Shirts, A4=Pants)
            # G2='Store LA' is at index 2 in B1:D1 (B1=Store NYC, C1=Store LA, D1=Store CHI)
            # INDEX(B2:D4, 2, 2) = C3 = 4500
            expected_value = ws.cell(row=3, column=3).value  # C3
            value_correct = (expected_value == 4500)

            if has_data_range and has_row_match and has_col_match and value_correct:
                print(f"PASS: Component 3 — Formula references correct ranges, expected result=4500 (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — data_range={has_data_range}, row_match={has_row_match}, col_match={has_col_match}, value_correct={value_correct}")
        else:
            print(f"FAIL: Component 3 — H2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
