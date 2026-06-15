"""
Initial Setup: Wide spreadsheet with 25 rows and 20 columns (A-T)
Task ID: calc_gfl_085
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_085'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # --- Headers (row 1, columns A-T) ---
    headers = [
        "ID", "Name", "Department", "Region", "Q1 Revenue", "Q2 Revenue",
        "Q3 Revenue", "Q4 Revenue", "Total Revenue", "Growth %",
        "Headcount", "Avg Salary", "Budget Alloc", "Expenses",
        "Net Profit", "Customer Count", "Satisfaction", "Returns",
        "Market Share %", "YoY Change %"
    ]
    header_font = Font(bold=True, size=11, name="Calibri")
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="000000")
    header_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11, name="Calibri", color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = header_border

    # --- Data rows (rows 2-26, 25 rows) ---
    names = [
        "Sarah Chen", "Marcus Johnson", "Priya Patel", "James O'Brien",
        "Yuki Tanaka", "Elena Volkov", "David Kim", "Amara Okafor",
        "Lucas Fernandez", "Sophie Martin", "Raj Gupta", "Emily Watson",
        "Carlos Rivera", "Fatima Al-Hassan", "Thomas Mueller",
        "Mei Lin Zhang", "Alexander Petrov", "Aisha Mohammed",
        "Ryan O'Connor", "Lena Johansson", "Hiroshi Nakamura",
        "Beatriz Costa", "Samuel Adeyemi", "Nadia Kowalski",
        "William Park"
    ]

    departments = [
        "Engineering", "Marketing", "Sales", "Operations", "Engineering",
        "Finance", "Product", "HR", "Sales", "Marketing",
        "Engineering", "Operations", "Sales", "Finance", "Product",
        "Engineering", "Marketing", "Sales", "Operations", "Finance",
        "Product", "Engineering", "HR", "Sales", "Marketing"
    ]

    regions = [
        "North America", "EMEA", "APAC", "North America", "APAC",
        "EMEA", "APAC", "North America", "LATAM", "EMEA",
        "APAC", "North America", "LATAM", "EMEA", "EMEA",
        "APAC", "EMEA", "North America", "EMEA", "EMEA",
        "APAC", "LATAM", "North America", "APAC", "EMEA"
    ]

    import random
    random.seed(42)

    for i in range(25):
        row = i + 2
        row_id = 1001 + i

        q1 = round(random.uniform(120000, 890000), 2)
        q2 = round(random.uniform(130000, 920000), 2)
        q3 = round(random.uniform(110000, 950000), 2)
        q4 = round(random.uniform(140000, 900000), 2)
        total = round(q1 + q2 + q3 + q4, 2)
        growth = round(random.uniform(-8.5, 25.3), 1)
        headcount = random.randint(12, 185)
        avg_salary = round(random.uniform(55000, 145000), 0)
        budget = round(random.uniform(200000, 1500000), 2)
        expenses = round(random.uniform(180000, 1400000), 2)
        net_profit = round(total - expenses, 2)
        customers = random.randint(150, 12000)
        satisfaction = round(random.uniform(3.2, 4.9), 1)
        returns = random.randint(5, 320)
        market_share = round(random.uniform(1.5, 18.7), 1)
        yoy_change = round(random.uniform(-12.0, 30.0), 1)

        row_data = [
            row_id, names[i], departments[i], regions[i],
            q1, q2, q3, q4, total, growth,
            headcount, avg_salary, budget, expenses,
            net_profit, customers, satisfaction, returns,
            market_share, yoy_change
        ]

        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.border = header_border
            # Format currency columns (E-I, N-O)
            if c in (5, 6, 7, 8, 9, 12, 13, 14, 15):
                cell.number_format = '$#,##0.00'
            # Format percentage columns (J, Q, S, T)
            elif c in (10, 17, 19, 20):
                cell.number_format = '0.0'
            # Format integer columns (K, P, R)
            elif c in (11, 16, 18):
                cell.number_format = '#,##0'

    # --- Column widths to make the sheet realistically wide ---
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    for col_letter in ["E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 16
    ws.column_dimensions["J"].width = 12
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 16
    ws.column_dimensions["N"].width = 16
    ws.column_dimensions["O"].width = 16
    ws.column_dimensions["P"].width = 16
    ws.column_dimensions["Q"].width = 14
    ws.column_dimensions["R"].width = 10
    ws.column_dimensions["S"].width = 16
    ws.column_dimensions["T"].width = 16

    # Freeze top row for easier navigation
    ws.freeze_panes = "A2"

    # NO print_title_cols set - this is what the task asks the agent to do
    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
