"""
Reward Script: Marketing Campaign ROI Tracker
Task ID: calc_grs_054
Domain: libreoffice_calc
Scoring:
  Component 1 (0.25): Formula columns I (Cost Per Lead), J (Cost Per Opportunity), K (Revenue ROI %)
  Component 2 (0.15): Conditional formatting on ROI % column (gold >200%, green >0%, red <0%)
  Component 3 (0.15): Scatter chart plotting Budget Spent vs Revenue Influenced
  Component 4 (0.25): Channel Summary sheet populated with SUMIF aggregation formulas
  Component 5 (0.20): Data sorted by ROI % descending
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_054'


def persist_app_state(domain: str):
    """Try to save any open LibreOffice file before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'Campaign Data' sheet must exist
    if 'Campaign Data' not in wb.sheetnames:
        print("CRITICAL: 'Campaign Data' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Campaign Data']

    # =========================================================================
    # Component 1: Formula columns I, J, K (0.25 points)
    # Task adds Cost Per Lead (I), Cost Per Opportunity (J), Revenue ROI % (K)
    # These columns do NOT exist in the initial file (initial only has A-H).
    # =========================================================================
    try:
        formula_score = 0.0
        # Check that column I (Cost Per Lead) has formulas referencing E and F
        # Check that column J (Cost Per Opportunity) has formulas referencing E and G
        # Check that column K (Revenue ROI %) has formulas referencing H, E
        i_formulas_ok = 0
        j_formulas_ok = 0
        k_formulas_ok = 0

        for row in range(2, 12):
            # Column I: Cost Per Lead = Budget / Leads => references E and F
            i_val = ws.cell(row=row, column=9).value
            if i_val is not None and isinstance(i_val, str) and '=' in i_val:
                i_upper = i_val.upper().replace(' ', '')
                if 'E' in i_upper and 'F' in i_upper:
                    i_formulas_ok += 1

            # Column J: Cost Per Opportunity = Budget / Opportunities => references E and G
            j_val = ws.cell(row=row, column=10).value
            if j_val is not None and isinstance(j_val, str) and '=' in j_val:
                j_upper = j_val.upper().replace(' ', '')
                if 'E' in j_upper and 'G' in j_upper:
                    j_formulas_ok += 1

            # Column K: Revenue ROI % = (Revenue - Budget) / Budget * 100 => references H and E
            k_val = ws.cell(row=row, column=11).value
            if k_val is not None and isinstance(k_val, str) and '=' in k_val:
                k_upper = k_val.upper().replace(' ', '')
                if 'H' in k_upper and 'E' in k_upper:
                    k_formulas_ok += 1

        # Need at least 8 out of 10 for each column to pass
        i_pass = i_formulas_ok >= 8
        j_pass = j_formulas_ok >= 8
        k_pass = k_formulas_ok >= 8

        cols_passing = sum([i_pass, j_pass, k_pass])
        if cols_passing == 3:
            formula_score = 0.25
            print(f"PASS: Component 1 -- All 3 formula columns present (I:{i_formulas_ok}/10, J:{j_formulas_ok}/10, K:{k_formulas_ok}/10) (0.25 pts)")
        elif cols_passing >= 1:
            formula_score = round(0.25 * cols_passing / 3, 2)
            print(f"PARTIAL: Component 1 -- {cols_passing}/3 formula columns (I:{i_formulas_ok}, J:{j_formulas_ok}, K:{k_formulas_ok}) ({formula_score} pts)")
        else:
            print(f"FAIL: Component 1 -- No formula columns found (I:{i_formulas_ok}, J:{j_formulas_ok}, K:{k_formulas_ok})")

        total_score += formula_score
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # =========================================================================
    # Component 2: Conditional formatting on ROI column (0.15 points)
    # Task requires: positive ROI in green, negative ROI in red, ROI above 200% in gold
    # Initial file has NO conditional formatting.
    # =========================================================================
    try:
        cf_score = 0.0
        cf_rules_found = 0
        has_green_rule = False
        has_red_rule = False
        has_gold_rule = False

        for cf in ws.conditional_formatting:
            cf_range = str(cf)
            # Check if formatting applies to column K area
            if 'K' in cf_range or '11' in cf_range:
                for rule in cf.rules:
                    rule_type = getattr(rule, 'type', '')
                    operator = getattr(rule, 'operator', '')
                    formula = getattr(rule, 'formula', [])
                    dxf = getattr(rule, 'dxf', None)

                    # Extract fill color from the rule
                    fill_rgb = None
                    if dxf and dxf.fill and dxf.fill.fgColor:
                        try:
                            fill_rgb = dxf.fill.fgColor.rgb
                        except:
                            pass

                    # Check for gold (>200%)
                    if operator in ('greaterThan', 'greaterThanOrEqual') and formula:
                        formula_str = str(formula[0]) if formula else ''
                        if '200' in formula_str:
                            has_gold_rule = True
                            cf_rules_found += 1

                    # Check for green (>0, positive ROI)
                    if operator in ('greaterThan', 'greaterThanOrEqual') and formula:
                        formula_str = str(formula[0]) if formula else ''
                        if formula_str.strip() == '0':
                            has_green_rule = True
                            cf_rules_found += 1

                    # Check for red (<0, negative ROI)
                    if operator in ('lessThan', 'lessThanOrEqual') and formula:
                        formula_str = str(formula[0]) if formula else ''
                        if formula_str.strip() == '0':
                            has_red_rule = True
                            cf_rules_found += 1

        rules_present = sum([has_green_rule, has_red_rule, has_gold_rule])
        if rules_present >= 3:
            cf_score = 0.15
            print(f"PASS: Component 2 -- All 3 conditional formatting rules found (green={has_green_rule}, red={has_red_rule}, gold={has_gold_rule}) (0.15 pts)")
        elif rules_present >= 1:
            cf_score = round(0.15 * rules_present / 3, 2)
            print(f"PARTIAL: Component 2 -- {rules_present}/3 CF rules (green={has_green_rule}, red={has_red_rule}, gold={has_gold_rule}) ({cf_score} pts)")
        else:
            print(f"FAIL: Component 2 -- No conditional formatting rules found on ROI column")

        total_score += cf_score
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # =========================================================================
    # Component 3: Scatter chart (0.15 points)
    # Task requires a scatter chart plotting Budget Spent vs Revenue Influenced.
    # Initial file has NO charts.
    # =========================================================================
    try:
        chart_score = 0.0
        charts_found = []

        # Check all sheets for scatter charts
        for sname in wb.sheetnames:
            sheet = wb[sname]
            for chart in sheet._charts:
                charts_found.append(chart)

        scatter_found = False
        for chart in charts_found:
            chart_type_name = type(chart).__name__
            if 'Scatter' in chart_type_name:
                scatter_found = True
                chart_score = 0.15
                print(f"PASS: Component 3 -- Scatter chart found (type={chart_type_name}, series={len(chart.series)}) (0.15 pts)")
                break

        if not scatter_found:
            # Accept any chart as partial credit (0.08)
            if len(charts_found) > 0:
                chart_score = 0.08
                ct = type(charts_found[0]).__name__
                print(f"PARTIAL: Component 3 -- Chart found but not ScatterChart (type={ct}) ({chart_score} pts)")
            else:
                print(f"FAIL: Component 3 -- No charts found in any sheet")

        total_score += chart_score
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # =========================================================================
    # Component 4: Channel Summary sheet with SUMIF aggregation (0.25 points)
    # Task requires a channel summary table using SUMIF showing aggregate performance by channel.
    # Initial file has 'Channel Summary' sheet with only a title in A1 -- no data.
    # Golden has 7 channels with SUMIF formulas.
    # =========================================================================
    try:
        summary_score = 0.0

        if 'Channel Summary' not in wb.sheetnames:
            print(f"FAIL: Component 4 -- 'Channel Summary' sheet not found")
        else:
            ws_summary = wb['Channel Summary']
            # Count cells containing SUMIF formulas
            sumif_count = 0
            channel_rows = 0

            # Scan reasonable range for SUMIF formulas
            for row in ws_summary.iter_rows(min_row=2, max_row=30, min_col=1, max_col=10, values_only=False):
                row_has_sumif = False
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        val_upper = cell.value.upper().replace(' ', '')
                        if 'SUMIF' in val_upper:
                            sumif_count += 1
                            row_has_sumif = True
                if row_has_sumif:
                    channel_rows += 1

            # Check for channel labels
            channel_labels = set()
            expected_channels = {'email', 'social', 'ppc', 'seo', 'content', 'events', 'pr'}
            for row in ws_summary.iter_rows(min_row=2, max_row=30, min_col=1, max_col=1, values_only=False):
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        label = cell.value.strip().lower()
                        if label in expected_channels:
                            channel_labels.add(label)

            channels_found = len(channel_labels)

            if sumif_count >= 20 and channels_found >= 5:
                # Full pass: many SUMIF formulas and most channels present
                summary_score = 0.25
                print(f"PASS: Component 4 -- Channel Summary has {sumif_count} SUMIF formulas across {channels_found} channels (0.25 pts)")
            elif sumif_count >= 10 and channels_found >= 3:
                summary_score = 0.18
                print(f"PARTIAL: Component 4 -- {sumif_count} SUMIFs, {channels_found} channels ({summary_score} pts)")
            elif sumif_count >= 4 or channels_found >= 2:
                summary_score = 0.10
                print(f"PARTIAL: Component 4 -- {sumif_count} SUMIFs, {channels_found} channels ({summary_score} pts)")
            else:
                print(f"FAIL: Component 4 -- Insufficient SUMIF data (SUMIFs={sumif_count}, channels={channels_found})")

        total_score += summary_score
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # =========================================================================
    # Component 5: Data sorted by ROI % descending (0.20 points)
    # Task requires sorting by ROI % descending.
    # Initial file is NOT sorted by ROI (different row order).
    # We verify by computing ROI from raw Budget (col E) and Revenue (col H) values.
    # =========================================================================
    try:
        sort_score = 0.0
        rois = []
        for row in range(2, ws.max_row + 1):
            budget = ws.cell(row=row, column=5).value
            revenue = ws.cell(row=row, column=8).value
            if budget and revenue and isinstance(budget, (int, float)) and isinstance(revenue, (int, float)) and budget > 0:
                roi = (revenue - budget) / budget * 100
                rois.append(roi)

        if len(rois) >= 8:
            # Check if fully sorted descending — must be true for ALL adjacent pairs
            is_sorted_desc = all(rois[i] >= rois[i + 1] for i in range(len(rois) - 1))
            if is_sorted_desc:
                sort_score = 0.20
                print(f"PASS: Component 5 -- Data sorted by ROI % descending ({len(rois)} rows, top={rois[0]:.1f}%, bottom={rois[-1]:.1f}%) (0.20 pts)")
            else:
                # Count how many adjacent pairs are in correct order
                correct_pairs = sum(1 for i in range(len(rois) - 1) if rois[i] >= rois[i + 1])
                total_pairs = len(rois) - 1
                # Only award partial if at least 70% of pairs are correctly ordered
                if correct_pairs / total_pairs >= 0.7:
                    sort_score = 0.10
                    print(f"PARTIAL: Component 5 -- Mostly sorted ({correct_pairs}/{total_pairs} pairs correct) ({sort_score} pts)")
                else:
                    print(f"FAIL: Component 5 -- Data NOT sorted by ROI % descending ({correct_pairs}/{total_pairs} pairs correct)")
                    for i, r in enumerate(rois[:5]):
                        print(f"  Row {i+2}: ROI={r:.1f}%")
        else:
            print(f"FAIL: Component 5 -- Not enough data rows to verify sorting ({len(rois)} rows found)")

        total_score += sort_score
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entrypoint
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
