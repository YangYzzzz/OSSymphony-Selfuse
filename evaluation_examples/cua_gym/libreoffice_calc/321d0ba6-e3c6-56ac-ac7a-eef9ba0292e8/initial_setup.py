"""
Initial Setup: Marketing Campaign ROI Tracker
Task ID: calc_grs_054
Domain: libreoffice_calc

Creates a spreadsheet with 10 marketing campaigns across multiple channels.
Contains raw data only — no formulas, no conditional formatting, no charts, unsorted.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

WORKDIR = '/home/user'
TASK_ID = 'calc_grs_054'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Campaign Data ---
    ws = wb.active
    ws.title = "Campaign Data"

    # Headers (only the raw data columns; formula columns will be added by the agent)
    headers = [
        "Campaign Name", "Channel", "Start Date", "End Date",
        "Budget Spent", "Leads Generated", "Opportunities Created",
        "Revenue Influenced"
    ]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align

    # 10 realistic marketing campaigns (NOT sorted by ROI)
    campaigns = [
        ["Spring Email Blast", "Email", "2025-01-15", "2025-02-28", 12500, 340, 45, 89000],
        ["LinkedIn Thought Leadership", "Social", "2025-02-01", "2025-04-30", 8200, 185, 22, 41500],
        ["Google Ads - Product Launch", "PPC", "2025-03-01", "2025-03-31", 35000, 920, 78, 156000],
        ["SEO Content Overhaul", "SEO", "2025-01-01", "2025-06-30", 18000, 1250, 95, 28000],
        ["Annual Marketing Summit", "Events", "2025-04-10", "2025-04-12", 52000, 280, 64, 187000],
        ["Blog Series - AI Trends", "Content", "2025-02-15", "2025-05-15", 6800, 410, 31, 15200],
        ["Industry Press Release", "PR", "2025-03-20", "2025-03-25", 4500, 95, 12, 8200],
        ["Instagram Reels Campaign", "Social", "2025-04-01", "2025-05-31", 15000, 520, 38, 22000],
        ["Retargeting PPC Funnel", "PPC", "2025-02-10", "2025-04-10", 22000, 680, 52, 98000],
        ["Webinar Lead Nurture", "Email", "2025-03-05", "2025-04-05", 9800, 290, 41, 67500],
    ]

    for r, row_data in enumerate(campaigns, 2):
        ws.cell(row=r, column=1, value=row_data[0])  # Campaign Name
        ws.cell(row=r, column=2, value=row_data[1])  # Channel
        ws.cell(row=r, column=3, value=row_data[2])  # Start Date
        ws.cell(row=r, column=4, value=row_data[3])  # End Date
        ws.cell(row=r, column=5, value=row_data[4])  # Budget Spent
        ws.cell(row=r, column=6, value=row_data[5])  # Leads Generated
        ws.cell(row=r, column=7, value=row_data[6])  # Opportunities Created
        ws.cell(row=r, column=8, value=row_data[7])  # Revenue Influenced

    # Format currency columns
    for r in range(2, 12):
        ws.cell(row=r, column=5).number_format = '$#,##0'
        ws.cell(row=r, column=8).number_format = '$#,##0'

    # Format date columns
    for r in range(2, 12):
        ws.cell(row=r, column=3).number_format = 'yyyy-mm-dd'
        ws.cell(row=r, column=4).number_format = 'yyyy-mm-dd'

    # Add data validation dropdown for Channel column
    dv = DataValidation(
        type="list",
        formula1='"Email,Social,PPC,SEO,Content,Events,PR"',
        allow_blank=False,
        showDropDown=False,  # False = show dropdown arrow
    )
    dv.error = "Please select a valid channel"
    dv.errorTitle = "Invalid Channel"
    dv.prompt = "Select marketing channel"
    dv.promptTitle = "Channel"
    dv.add("B2:B11")
    ws.add_data_validation(dv)

    # Set column widths for readability
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 20

    # Freeze header row
    ws.freeze_panes = "A2"

    # --- Sheet 2: Channel Summary (empty placeholder — agent will build SUMIF table) ---
    ws2 = wb.create_sheet("Channel Summary")
    ws2["A1"] = "Channel Summary"
    ws2["A1"].font = Font(name="Calibri", size=14, bold=True)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
