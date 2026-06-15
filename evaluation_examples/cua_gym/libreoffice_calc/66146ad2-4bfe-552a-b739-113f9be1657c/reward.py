"""
Reward Script: Format appointment times in column C to 12-hour AM/PM format
Task ID: calc_fmt_numfmt_time_ampm_090
Domain: libreoffice_calc
Scoring:
  Component 1 (0.6): All cells C2:C30 have number_format 'H:MM AM/PM'
  Component 2 (0.4): All cells C2:C30 contain valid datetime.time values (not float fractions)
"""

import os
import datetime
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_fmt_numfmt_time_ampm_090'

TARGET_FORMAT = 'H:MM AM/PM'
DATA_ROWS_START = 2
DATA_ROWS_END = 30  # inclusive
TOTAL_DATA_CELLS = DATA_ROWS_END - DATA_ROWS_START + 1  # 29


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0

    Task: Format cells C2:C30 in the 'Appointment Book' sheet to display
    in 12-hour format with AM/PM (e.g., '9:30 AM', '2:15 PM').

    The initial file has column C with 'General' format and raw float fractions.
    The golden file should have 'H:MM AM/PM' format and datetime.time values.
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition gate: 'Appointment Book' sheet must exist
    if 'Appointment Book' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Appointment Book' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Appointment Book']

    # Precondition gate: C1 header should be 'Time' — if missing, file is corrupted
    c1_val = ws.cell(row=1, column=3).value
    if c1_val != 'Time':
        print(f"CRITICAL: C1 header expected 'Time', found {c1_val!r}. File appears corrupted.")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: All cells C2:C30 have number_format 'H:MM AM/PM' (0.6 points)
    # Initial file: all cells have 'General' format → FAILS on initial
    # Golden file: all cells have 'H:MM AM/PM' format → PASSES on golden
    try:
        formatted_count = 0

        for row in range(DATA_ROWS_START, DATA_ROWS_END + 1):
            cell = ws.cell(row=row, column=3)
            if cell.number_format == TARGET_FORMAT:
                formatted_count += 1

        if formatted_count == TOTAL_DATA_CELLS:
            print(f"PASS: Component 1 — All {TOTAL_DATA_CELLS} cells C2:C30 have number_format '{TARGET_FORMAT}' (0.6 pts)")
            total_score += 0.6
        elif formatted_count > 0:
            partial = round(0.6 * formatted_count / TOTAL_DATA_CELLS, 4)
            print(f"PARTIAL: Component 1 — {formatted_count}/{TOTAL_DATA_CELLS} cells have '{TARGET_FORMAT}' format (partial {partial} pts)")
            total_score += partial
        else:
            first_fmt = ws.cell(row=2, column=3).number_format
            print(f"FAIL: Component 1 — 0/{TOTAL_DATA_CELLS} cells have '{TARGET_FORMAT}' format; C2 has format '{first_fmt}'")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C2:C30 contain valid datetime.time values (not raw float fractions) (0.4 points)
    # Initial file: cells contain float fractions (e.g., 0.395833) → FAILS on initial
    # Golden file: cells contain datetime.time objects (e.g., datetime.time(9, 30)) → PASSES on golden
    # Note: openpyxl reads time-formatted cells as datetime.time objects, not floats.
    try:
        time_val_count = 0
        for row in range(DATA_ROWS_START, DATA_ROWS_END + 1):
            cell = ws.cell(row=row, column=3)
            val = cell.value
            if isinstance(val, datetime.time):
                time_val_count += 1

        if time_val_count == TOTAL_DATA_CELLS:
            print(f"PASS: Component 2 — All {TOTAL_DATA_CELLS} cells C2:C30 contain datetime.time values (0.4 pts)")
            total_score += 0.4
        elif time_val_count > 0:
            partial = round(0.4 * time_val_count / TOTAL_DATA_CELLS, 4)
            print(f"PARTIAL: Component 2 — {time_val_count}/{TOTAL_DATA_CELLS} cells have datetime.time values (partial {partial} pts)")
            total_score += partial
        else:
            first_val = ws.cell(row=2, column=3).value
            print(f"FAIL: Component 2 — 0/{TOTAL_DATA_CELLS} cells have datetime.time values; C2 value={first_val!r} (type={type(first_val).__name__})")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    final_score = min(round(total_score, 4), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
