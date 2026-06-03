"""
Initial Setup: Appointment Book spreadsheet with time values in General format
Task ID: calc_fmt_numfmt_time_ampm_090
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

WORKDIR = '/home/user'
TASK_ID = 'calc_fmt_numfmt_time_ampm_090'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Appointment Book'

    # --- Headers (Row 1) ---
    headers = ['Client', 'Date', 'Time', 'Service']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # --- Appointment data (Rows 2-30) ---
    # Time fractions: value = (hour + minute/60) / 24
    # e.g., 9:30 AM = (9 + 30/60) / 24 = 9.5/24 = 0.395833...
    # e.g., 14:15 PM = (14 + 15/60) / 24 = 14.25/24 = 0.593750
    appointments = [
        ('Emily Hartman',     '2025-06-02', 9.5/24,          'Haircut & Style'),
        ('Marcus Webb',       '2025-06-02', 10.0/24,         'Deep Tissue Massage'),
        ('Sophia Ramirez',    '2025-06-02', 10.75/24,        'Manicure'),
        ('Daniel Osei',       '2025-06-02', 11.5/24,         'Facial Treatment'),
        ('Priya Nair',        '2025-06-02', 12.25/24,        'Pedicure'),
        ('James Thornton',    '2025-06-03', 9.0/24,          'Haircut'),
        ('Ava Kowalski',      '2025-06-03', 9.75/24,         'Color & Highlights'),
        ('Ethan Fitzgerald',  '2025-06-03', 11.0/24,         'Sports Massage'),
        ('Lily Chen',         '2025-06-03', 13.5/24,         'Gel Nails'),
        ('Noah Patel',        '2025-06-03', 14.25/24,        'Back Massage'),
        ('Chloe Andersen',    '2025-06-04', 8.5/24,          'Eyebrow Threading'),
        ('Lucas Moreau',      '2025-06-04', 10.5/24,         'Haircut & Beard'),
        ('Isabelle Fontaine', '2025-06-04', 11.25/24,        'Deep Cleanse Facial'),
        ('Aiden Brooks',      '2025-06-04', 13.0/24,         'Foot Massage'),
        ('Maya Johansson',    '2025-06-04', 15.0/24,         'Waxing'),
        ('Liam Castro',       '2025-06-05', 9.25/24,         'Haircut'),
        ('Nadia Petrov',      '2025-06-05', 10.0/24,         'Shellac Manicure'),
        ('Ryan O\'Brien',     '2025-06-05', 11.75/24,        'Swedish Massage'),
        ('Simone Lefebvre',   '2025-06-05', 13.25/24,        'Hair Coloring'),
        ('Victor Huang',      '2025-06-05', 14.5/24,         'Facial'),
        ('Grace Williamson',  '2025-06-06', 9.5/24,          'Pedicure & Massage'),
        ('Owen Nakamura',     '2025-06-06', 11.0/24,         'Haircut'),
        ('Zoe Fernandez',     '2025-06-06', 12.5/24,         'Bridal Makeup'),
        ('Henry Larsson',     '2025-06-06', 14.0/24,         'Hot Stone Massage'),
        ('Amelia Koh',        '2025-06-06', 15.5/24,         'Nail Art'),
        ('Jack Tremblay',     '2025-06-07', 9.0/24,          'Haircut & Shave'),
        ('Leila Mansouri',    '2025-06-07', 10.25/24,        'Lash Extensions'),
        ('Samuel Diaz',       '2025-06-07', 13.75/24,        'Deep Tissue Massage'),
        ('Clara Bergmann',    '2025-06-07', 15.25/24,        'Manicure & Pedicure'),
    ]

    for r, (client, date_str, time_frac, service) in enumerate(appointments, 2):
        ws.cell(row=r, column=1, value=client)
        ws.cell(row=r, column=2, value=date_str)
        # Column C: time fraction value, General format (NOT formatted as time yet)
        time_cell = ws.cell(row=r, column=3, value=time_frac)
        time_cell.number_format = 'General'
        ws.cell(row=r, column=4, value=service)

    # Column widths for readability
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 25

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  Rows: 30 (header + 29 appointments)')
    print(f'  Column C uses General format (task not yet applied)')


create_initial()
