"""
Initial Setup: External reference with stale cached value
Task ID: calc_tbl_089
Domain: libreoffice_calc

Creates a workbook with an external reference formula pointing to /root/data/Prices.xlsx.
The formula caches 29.99 but the external file has been updated to 34.99.
Also creates the external Prices.xlsx with the updated value.
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_089'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'
EXTERNAL_DIR = '/home/user/data'
EXTERNAL_FILE = f'{EXTERNAL_DIR}/Prices.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_external_file():
    """Create the external Prices.xlsx with the UPDATED value 34.99."""
    os.makedirs(EXTERNAL_DIR, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    ws['A1'] = 34.99  # Updated price (was 29.99, now 34.99)
    ws['B1'] = 'Premium Widget'

    # Additional price data for realism
    headers = ['Price', 'Product Name', 'SKU', 'Category', 'Last Updated']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h if col > 1 else 34.99)

    # Wait - A1 should just be the price value. Let me restructure.
    # Row 1 is the data row referenced by the external link
    ws['A1'] = 34.99
    ws['B1'] = 'Premium Widget'
    ws['C1'] = 'SKU-4821'
    ws['D1'] = 'Electronics'
    ws['E1'] = '2026-03-28'

    # More product prices
    products = [
        [22.50, 'Standard Widget', 'SKU-4822', 'Electronics', '2026-03-28'],
        [15.99, 'Basic Widget', 'SKU-4823', 'Electronics', '2026-03-15'],
        [89.99, 'Deluxe Widget', 'SKU-4824', 'Electronics', '2026-03-20'],
        [45.00, 'Professional Widget', 'SKU-4825', 'Electronics', '2026-03-22'],
    ]
    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    wb.save(EXTERNAL_FILE)
    print(f'External file created: {EXTERNAL_FILE} with A1=34.99')


def create_initial():
    """Create the main workbook with an external reference showing stale cached value."""
    wb = openpyxl.Workbook()

    # --- Sheet 1: Inventory ---
    ws1 = wb.active
    ws1.title = 'Inventory'

    headers = ['Product', 'Unit Price', 'Quantity', 'Total Value', 'Supplier', 'Reorder Level']
    header_font = Font(bold=True, size=11, name='Calibri')
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(bold=True, size=11, name='Calibri', color='FFFFFF')

    # Data rows - realistic inventory
    inventory_data = [
        ['Premium Widget',    None,   150, None, 'TechParts Co.',     50],
        ['Standard Widget',   22.50,  300, None, 'TechParts Co.',     100],
        ['Basic Widget',      15.99,  500, None, 'ValueSupply Inc.',  200],
        ['Deluxe Widget',     89.99,   45, None, 'TechParts Co.',     20],
        ['Professional Widget', 45.00, 120, None, 'ProGear Ltd.',     40],
        ['Economy Widget',    12.50,  800, None, 'ValueSupply Inc.',  300],
        ['Mini Widget',        8.99,  650, None, 'SmallParts Corp.',  250],
        ['Mega Widget',      125.00,   30, None, 'ProGear Ltd.',      10],
        ['Ultra Widget',      67.50,   75, None, 'TechParts Co.',     25],
        ['Compact Widget',    19.99,  400, None, 'SmallParts Corp.',  150],
        ['Travel Widget',     35.00,  200, None, 'ProGear Ltd.',      60],
        ['Home Widget',       28.75,  350, None, 'ValueSupply Inc.',  120],
    ]

    for r, row_data in enumerate(inventory_data, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # Row 2, Column B (Unit Price for Premium Widget) - external reference with STALE cached value
    # The formula references the external file, but we write it as a formula string.
    # openpyxl stores the formula; the cached value 29.99 is what shows until refreshed.
    ext_ref = "='[/home/user/data/Prices.xlsx]Sheet1'.A1"
    ws1.cell(row=2, column=2, value=ext_ref)

    # Total Value formulas for all rows
    for r in range(2, 14):
        ws1.cell(row=r, column=4, value=f'=B{r}*C{r}')

    # Summary row
    ws1.cell(row=15, column=1, value='TOTAL')
    ws1.cell(row=15, column=1).font = Font(bold=True, size=11)
    ws1.cell(row=15, column=4, value='=SUM(D2:D13)')
    ws1.cell(row=15, column=4).font = Font(bold=True, size=11)

    # Column widths
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 12
    ws1.column_dimensions['D'].width = 14
    ws1.column_dimensions['E'].width = 20
    ws1.column_dimensions['F'].width = 14

    # Freeze header row
    ws1.freeze_panes = 'A2'

    # --- Sheet 2: Suppliers ---
    ws2 = wb.create_sheet('Suppliers')

    supplier_headers = ['Company', 'Contact', 'Phone', 'Email', 'Region', 'Payment Terms']
    for col, h in enumerate(supplier_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)

    suppliers = [
        ['TechParts Co.',     'Diana Ross',    '555-0101', 'diana@techparts.com',     'Northeast', 'Net 30'],
        ['ValueSupply Inc.',  'Robert Chen',   '555-0202', 'robert@valuesupply.com',  'Midwest',   'Net 45'],
        ['ProGear Ltd.',      'Amanda White',  '555-0303', 'amanda@progear.com',      'West Coast','Net 30'],
        ['SmallParts Corp.',  'Kevin Park',    '555-0404', 'kevin@smallparts.com',    'Southeast', 'Net 60'],
    ]

    for r, row_data in enumerate(suppliers, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 20
    ws2.column_dimensions['B'].width = 16
    ws2.column_dimensions['C'].width = 12
    ws2.column_dimensions['D'].width = 28
    ws2.column_dimensions['E'].width = 14
    ws2.column_dimensions['F'].width = 16

    # --- Sheet 3: Order History ---
    ws3 = wb.create_sheet('Order History')

    order_headers = ['Order ID', 'Date', 'Product', 'Qty', 'Unit Price', 'Total', 'Status']
    for col, h in enumerate(order_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)

    orders = [
        ['ORD-2026-001', '2026-01-10', 'Premium Widget',      50,  29.99, 1499.50, 'Delivered'],
        ['ORD-2026-002', '2026-01-18', 'Basic Widget',        200,  15.99, 3198.00, 'Delivered'],
        ['ORD-2026-003', '2026-02-05', 'Deluxe Widget',        10,  89.99,  899.90, 'Delivered'],
        ['ORD-2026-004', '2026-02-14', 'Standard Widget',     100,  22.50, 2250.00, 'Delivered'],
        ['ORD-2026-005', '2026-02-28', 'Professional Widget',  30,  45.00, 1350.00, 'In Transit'],
        ['ORD-2026-006', '2026-03-05', 'Premium Widget',       25,  29.99,  749.75, 'In Transit'],
        ['ORD-2026-007', '2026-03-12', 'Economy Widget',      300,  12.50, 3750.00, 'Processing'],
        ['ORD-2026-008', '2026-03-20', 'Mega Widget',           5, 125.00,  625.00, 'Processing'],
    ]

    for r, row_data in enumerate(orders, 2):
        for c, val in enumerate(row_data, 1):
            ws3.cell(row=r, column=c, value=val)

    ws3.column_dimensions['A'].width = 16
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 22
    ws3.column_dimensions['D'].width = 8
    ws3.column_dimensions['E'].width = 12
    ws3.column_dimensions['F'].width = 12
    ws3.column_dimensions['G'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # Open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


# Execute
create_external_file()
create_initial()
