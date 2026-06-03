"""
Reward Script: Refresh external workbook link in cell B2
Task ID: calc_tbl_089
Domain: libreoffice_calc
Scoring:
  Component 1 (0.5 pts): B2 displays the current value 34.99 (not old cached 29.99)
  Component 2 (0.3 pts): B2 no longer contains external reference formula
  Component 3 (0.2 pts): Remaining data integrity - other prices and formulas intact
"""

import os
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_tbl_089'


def persist_app_state(domain: str):
    """Save any unsaved GUI edits before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Load workbook in formula mode to check if B2 still has external ref
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Inventory']

    # Also load in data_only mode to get cached computed values
    try:
        wb_data = openpyxl.load_workbook(file_path, data_only=True)
        ws_data = wb_data['Inventory']
    except Exception as e:
        print(f"WARN: Cannot load data_only mode: {e}")
        ws_data = None

    b2_formula_value = ws['B2'].value
    b2_data_value = ws_data['B2'].value if ws_data else None

    print(f"DEBUG: B2 formula-mode value: {b2_formula_value!r}")
    print(f"DEBUG: B2 data-only value: {b2_data_value!r}")

    # Determine the effective B2 value (either direct numeric or cached)
    effective_b2 = None
    if isinstance(b2_formula_value, (int, float)):
        effective_b2 = float(b2_formula_value)
    elif b2_data_value is not None and isinstance(b2_data_value, (int, float)):
        effective_b2 = float(b2_data_value)

    # Check if B2 still has an external reference formula
    b2_has_external_ref = (
        isinstance(b2_formula_value, str)
        and b2_formula_value.startswith('=')
        and '[' in b2_formula_value
    )

    # Component 1: B2 displays the current value 34.99 (0.5 points)
    # This is the core task requirement: the cell should show 34.99, not 29.99
    try:
        if effective_b2 is not None and abs(effective_b2 - 34.99) < 0.01:
            print(f"PASS: Component 1 -- B2 has correct value 34.99 (found: {effective_b2}) (0.5 pts)")
            total_score += 0.5
        else:
            print(f"FAIL: Component 1 -- Expected B2 ~= 34.99, found formula={b2_formula_value!r}, data={b2_data_value!r}, effective={effective_b2}")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: B2 no longer contains an external reference formula (0.3 points)
    # The task says to replace with the current value OR refresh the link.
    # In golden, the formula was replaced with static value 34.99.
    # We accept either: (a) static value 34.99, or (b) external ref formula
    # BUT only if the displayed value is 34.99 (checked in Component 1).
    # The key discriminator: in initial_env, B2 has the external ref AND cached value 29.99.
    # In golden_env, B2 has static value 34.99 (no formula).
    # We give 0.3 pts if the external reference is gone (replaced with static value).
    try:
        if not b2_has_external_ref:
            print(f"PASS: Component 2 -- B2 external reference removed (value: {b2_formula_value!r}) (0.3 pts)")
            total_score += 0.3
        else:
            # If the formula is still there but value is refreshed (34.99), give partial
            if effective_b2 is not None and abs(effective_b2 - 34.99) < 0.01:
                print(f"PARTIAL: Component 2 -- External ref still present but value refreshed, giving partial (0.15 pts)")
                total_score += 0.15
            else:
                print(f"FAIL: Component 2 -- B2 still has external reference: {b2_formula_value!r}")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: Data integrity - other prices and formulas intact (0.2 points)
    # Verify that the rest of the Inventory sheet was not corrupted
    # This is a compound check: B2 updated to 34.99 AND other prices unchanged
    try:
        expected_prices = {
            'B3': 22.5,
            'B4': 15.99,
            'B5': 89.99,
            'B6': 45,
            'B7': 12.5,
            'B8': 8.99,
            'B9': 125,
            'B10': 67.5,
            'B11': 19.99,
            'B12': 35,
            'B13': 28.75,
        }
        mismatches = []
        for coord, expected_val in expected_prices.items():
            cell_val = ws[coord].value
            if cell_val is None or not isinstance(cell_val, (int, float)):
                mismatches.append(f"{coord}: expected {expected_val}, found {cell_val!r}")
            elif abs(float(cell_val) - expected_val) > 0.01:
                mismatches.append(f"{coord}: expected {expected_val}, found {cell_val}")

        if len(mismatches) == 0 and effective_b2 is not None and abs(effective_b2 - 34.99) < 0.01:
            print(f"PASS: Component 3 -- Data integrity verified, other prices intact (0.2 pts)")
            total_score += 0.2
        elif len(mismatches) > 0:
            print(f"FAIL: Component 3 -- Data integrity compromised: {mismatches[0]}")
        else:
            print(f"FAIL: Component 3 -- B2 not updated, so integrity check anchored to task change fails")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Persist any unsaved GUI state before verification
persist_app_state("libreoffice_calc")

# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
