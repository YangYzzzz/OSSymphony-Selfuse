"""
Initial Setup: Create timesheet with headers and employee data, no validation
Task ID: calc_nrv_087
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_nrv_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timesheet"

    # Headers
    headers = ['Date', 'Employee', 'Regular Hours', 'Overtime Hours']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Realistic timesheet data: 30 rows of dates and employee names
    # C and D columns left EMPTY (no hours, no validation)
    employees = [
        'Sarah Chen', 'Marcus Johnson', 'Priya Patel', 'James O\'Brien',
        'Aisha Mohammed', 'Carlos Rivera', 'Emma Larsson', 'David Kim',
        'Fatima Al-Rashid', 'Lucas Fernandez', 'Mei-Lin Wu', 'Roberto Santos',
        'Anna Kowalski', 'Takeshi Yamamoto', 'Sofia Garcia',
    ]

    import datetime
    start_date = datetime.date(2025, 3, 1)
    for i in range(30):
        row = i + 2
        current_date = start_date + datetime.timedelta(days=i)
        ws.cell(row=row, column=1, value=current_date)
        ws.cell(row=row, column=1).number_format = 'yyyy-mm-dd'
        ws.cell(row=row, column=2, value=employees[i % len(employees)])
        # C2:C31 and D2:D31 intentionally left empty - no validation

    # Adjust column widths for readability
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 18

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready: open in LibreOffice Calc
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
