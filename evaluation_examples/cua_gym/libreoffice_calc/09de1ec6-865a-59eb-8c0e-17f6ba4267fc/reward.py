"""
Reward Script: Create a formatted energy bill analysis with monthly comparison and seasonal pattern chart.
Task ID: calc_gpm_081
Domain: libreoffice_calc
Scoring:
  Component 1: Title merge + styling (A1:G1)        — 0.10 pts
  Component 2: Header row 3 bold + styled            — 0.10 pts
  Component 3: Tiered rate formulas in C4:C15 (IF)   — 0.15 pts
  Component 4: Bill calc formulas D, F, G             — 0.20 pts
  Component 5: Summary rows 17-18 with SUM/AVERAGE   — 0.15 pts
  Component 6: Chart present with correct title       — 0.15 pts
  Component 7: Conditional formatting on G4:G15       — 0.10 pts
  Component 8: Number formats applied                 — 0.05 pts
  Total: 1.00
"""

import os

import openpyxl
from openpyxl.cell.cell import MergedCell

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_081'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify sheet 'Energy' exists
    if 'Energy' not in wb.sheetnames:
        print("CRITICAL: Sheet 'Energy' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Energy']

    # ---------------------------------------------------------------
    # Component 1: Title merge + styling in A1:G1 (0.10 points)
    # Initial: A1 has text but no merge, no bold, size 11
    # Golden: A1:G1 merged, bold, 14pt, dark green fill, white text, centered
    # ---------------------------------------------------------------
    try:
        a1 = ws['A1']
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        is_merged = any('A1' in str(mr) and 'G1' in str(mr) for mr in ws.merged_cells.ranges)
        is_bold = a1.font.bold is True
        is_14pt = a1.font.size is not None and abs(a1.font.size - 14) < 1
        is_centered = a1.alignment.horizontal == 'center'

        # Check dark green fill (RGB 0,100,0 = 006400)
        has_green_fill = False
        try:
            fill_rgb = a1.fill.fgColor.rgb
            if fill_rgb and '006400' in str(fill_rgb):
                has_green_fill = True
        except Exception:
            pass

        checks_passed = sum([is_merged, is_bold, is_14pt, is_centered, has_green_fill])
        if checks_passed >= 4:
            print(f"PASS: Component 1 — Title merge+styling ({checks_passed}/5 sub-checks) (0.10 pts)")
            total_score += 0.10
        elif checks_passed >= 2:
            partial = 0.05
            print(f"PARTIAL: Component 1 — Title ({checks_passed}/5 sub-checks) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 1 — Title merge+styling ({checks_passed}/5: merged={is_merged}, bold={is_bold}, 14pt={is_14pt}, centered={is_centered}, green_fill={has_green_fill})")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ---------------------------------------------------------------
    # Component 2: Header row 3 bold + styled (0.10 points)
    # Initial: headers exist but no bold, no fill, no borders
    # Golden: bold, green fill, white text, borders
    # ---------------------------------------------------------------
    try:
        bold_count = 0
        fill_count = 0
        border_count = 0
        for col in range(1, 8):
            cell = ws.cell(row=3, column=col)
            if cell.font.bold is True:
                bold_count += 1
            try:
                fill_rgb = cell.fill.fgColor.rgb
                if fill_rgb and cell.fill.patternType == 'solid' and fill_rgb != '00000000':
                    fill_count += 1
            except Exception:
                pass
            if cell.border.left.style is not None or cell.border.right.style is not None:
                border_count += 1

        if bold_count >= 6 and fill_count >= 5:
            print(f"PASS: Component 2 — Headers bold+styled (bold={bold_count}/7, fill={fill_count}/7, border={border_count}/7) (0.10 pts)")
            total_score += 0.10
        elif bold_count >= 5:
            print(f"PARTIAL: Component 2 — Headers bold but partial styling (bold={bold_count}, fill={fill_count}) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 2 — Headers not styled (bold={bold_count}/7, fill={fill_count}/7)")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ---------------------------------------------------------------
    # Component 3: Tiered rate formulas in C4:C15 (0.15 points)
    # Initial: C4:C15 all None
    # Golden: =IF(B4<=500,0.12,0.18) pattern
    # ---------------------------------------------------------------
    try:
        if_formula_count = 0
        for r in range(4, 16):
            val = ws.cell(row=r, column=3).value
            if val and isinstance(val, str) and 'IF' in val.upper() and '0.12' in str(val) and '0.18' in str(val):
                if_formula_count += 1

        if if_formula_count >= 11:
            print(f"PASS: Component 3 — Tiered rate IF formulas ({if_formula_count}/12) (0.15 pts)")
            total_score += 0.15
        elif if_formula_count >= 6:
            partial = 0.08
            print(f"PARTIAL: Component 3 — Some IF formulas ({if_formula_count}/12) ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — IF formulas in C column ({if_formula_count}/12)")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    # ---------------------------------------------------------------
    # Component 4: Bill calc formulas in D, F, G columns (0.20 points)
    # Initial: D, F, G all None
    # Golden: D=B*C, F=E*1.25, G=D+F
    # ---------------------------------------------------------------
    try:
        d_formula_count = 0
        f_formula_count = 0
        g_formula_count = 0

        for r in range(4, 16):
            # D column: Electric Bill = B*C
            d_val = ws.cell(row=r, column=4).value
            if d_val and isinstance(d_val, str) and '=' in d_val:
                d_upper = d_val.upper().replace(' ', '')
                if ('B' in d_upper and 'C' in d_upper) or 'B' in d_upper:
                    d_formula_count += 1

            # F column: Gas Bill = E*1.25
            f_val = ws.cell(row=r, column=6).value
            if f_val and isinstance(f_val, str) and '=' in f_val and '1.25' in f_val:
                f_formula_count += 1

            # G column: Total Bill = D+F
            g_val = ws.cell(row=r, column=7).value
            if g_val and isinstance(g_val, str) and '=' in g_val:
                g_upper = g_val.upper().replace(' ', '')
                if 'D' in g_upper and 'F' in g_upper:
                    g_formula_count += 1

        # Score: D formulas (0.07), F formulas (0.06), G formulas (0.07)
        sub_score = 0.0
        if d_formula_count >= 10:
            sub_score += 0.07
            print(f"  D column formulas: PASS ({d_formula_count}/12)")
        else:
            print(f"  D column formulas: FAIL ({d_formula_count}/12)")

        if f_formula_count >= 10:
            sub_score += 0.06
            print(f"  F column formulas: PASS ({f_formula_count}/12)")
        else:
            print(f"  F column formulas: FAIL ({f_formula_count}/12)")

        if g_formula_count >= 10:
            sub_score += 0.07
            print(f"  G column formulas: PASS ({g_formula_count}/12)")
        else:
            print(f"  G column formulas: FAIL ({g_formula_count}/12)")

        if sub_score > 0:
            print(f"PASS: Component 4 — Bill calc formulas ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 4 — No bill calc formulas found")
    except Exception as e:
        print(f"ERROR: Component 4 — {e}")

    # ---------------------------------------------------------------
    # Component 5: Summary rows 17-18 with SUM/AVERAGE (0.15 points)
    # Initial: rows 17-18 all None
    # Golden: Row 17 "Annual Summary" + SUM formulas, Row 18 "Monthly Average" + AVERAGE formulas
    # ---------------------------------------------------------------
    try:
        sub_score = 0.0

        # Row 17: Annual Summary label
        a17_val = ws.cell(row=17, column=1).value
        has_summary_label = a17_val and 'annual' in str(a17_val).lower() and 'summary' in str(a17_val).lower()

        # Row 17: SUM formulas in B17, D17, F17, G17
        sum_count = 0
        for col in [2, 4, 6, 7]:
            val = ws.cell(row=17, column=col).value
            if val and isinstance(val, str) and 'SUM' in val.upper():
                sum_count += 1

        # Row 18: Monthly Average label
        a18_val = ws.cell(row=18, column=1).value
        has_avg_label = a18_val and 'average' in str(a18_val).lower()

        # Row 18: AVERAGE formulas
        avg_count = 0
        for col in [2, 4, 6, 7]:
            val = ws.cell(row=18, column=col).value
            if val and isinstance(val, str) and 'AVERAGE' in val.upper():
                avg_count += 1

        if has_summary_label and sum_count >= 3:
            sub_score += 0.08
            print(f"  Row 17 Annual Summary: PASS (label={has_summary_label}, SUM formulas={sum_count}/4)")
        else:
            print(f"  Row 17 Annual Summary: FAIL (label={has_summary_label}, SUM formulas={sum_count}/4)")

        if has_avg_label and avg_count >= 3:
            sub_score += 0.07
            print(f"  Row 18 Monthly Average: PASS (label={has_avg_label}, AVG formulas={avg_count}/4)")
        else:
            print(f"  Row 18 Monthly Average: FAIL (label={has_avg_label}, AVG formulas={avg_count}/4)")

        if sub_score > 0:
            print(f"PASS: Component 5 — Summary rows ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 5 — No summary rows found")
    except Exception as e:
        print(f"ERROR: Component 5 — {e}")

    # ---------------------------------------------------------------
    # Component 6: Chart present with correct title (0.15 points)
    # Initial: 0 charts
    # Golden: 1 chart, col type, title "Monthly Energy Cost & Usage"
    # ---------------------------------------------------------------
    try:
        charts = ws._charts
        if len(charts) >= 1:
            chart = charts[0]
            # Check chart title
            has_title = False
            title_text = ''
            try:
                if chart.title:
                    # Extract title text from the rich text structure
                    for para in chart.title.tx.rich.paragraphs:
                        for run in para.r:
                            title_text += run.t
                    if 'energy' in title_text.lower() and ('cost' in title_text.lower() or 'usage' in title_text.lower()):
                        has_title = True
            except Exception:
                pass

            if has_title:
                print(f"PASS: Component 6 — Chart present with title '{title_text}' (0.15 pts)")
                total_score += 0.15
            else:
                print(f"PARTIAL: Component 6 — Chart present but title mismatch (title='{title_text}') (0.08 pts)")
                total_score += 0.08
        else:
            print(f"FAIL: Component 6 — No charts found ({len(charts)} charts)")
    except Exception as e:
        print(f"ERROR: Component 6 — {e}")

    # ---------------------------------------------------------------
    # Component 7: Conditional formatting on G4:G15 (0.10 points)
    # Initial: no conditional formatting
    # Golden: 3 rules on G4:G15 (>300 red, 200-300 orange, <200 green)
    #         + color scale on B4:B15
    # ---------------------------------------------------------------
    try:
        cf_on_g = False
        cf_rules_count = 0
        cf_on_b = False

        for cf in ws.conditional_formatting:
            cf_str = str(cf)
            # Check for G column conditional formatting
            if 'G4' in cf_str or 'G15' in cf_str:
                cf_on_g = True
                cf_rules_count = len(cf.rules)
            # Check for B column color scale
            if 'B4' in cf_str or 'B15' in cf_str:
                for rule in cf.rules:
                    if rule.type == 'colorScale':
                        cf_on_b = True

        sub_score = 0.0
        if cf_on_g and cf_rules_count >= 2:
            sub_score += 0.07
            print(f"  G column CF: PASS ({cf_rules_count} rules)")
        else:
            print(f"  G column CF: FAIL (found={cf_on_g}, rules={cf_rules_count})")

        if cf_on_b:
            sub_score += 0.03
            print(f"  B column color scale: PASS")
        else:
            print(f"  B column color scale: FAIL")

        if sub_score > 0:
            print(f"PASS: Component 7 — Conditional formatting ({sub_score} pts)")
            total_score += sub_score
        else:
            print(f"FAIL: Component 7 — No conditional formatting found")
    except Exception as e:
        print(f"ERROR: Component 7 — {e}")

    # ---------------------------------------------------------------
    # Component 8: Number formats applied (0.05 points)
    # Initial: all General format
    # Golden: C=$0.000, D/G=$#,##0.00, B/E=#,##0
    # ---------------------------------------------------------------
    try:
        correct_formats = 0

        # C column should have $0.000 (currency with 3 decimals)
        c_fmt = ws['C4'].number_format
        if c_fmt and '$' in c_fmt and '0' in c_fmt and c_fmt != 'General':
            correct_formats += 1

        # D column should have $#,##0.00
        d_fmt = ws['D4'].number_format
        if d_fmt and '$' in d_fmt and d_fmt != 'General':
            correct_formats += 1

        # B column should have #,##0
        b_fmt = ws['B4'].number_format
        if b_fmt and '#' in b_fmt and ',' in b_fmt and b_fmt != 'General':
            correct_formats += 1

        # G column should have $#,##0.00
        g_fmt = ws['G4'].number_format
        if g_fmt and '$' in g_fmt and g_fmt != 'General':
            correct_formats += 1

        if correct_formats >= 3:
            print(f"PASS: Component 8 — Number formats ({correct_formats}/4 columns correct) (0.05 pts)")
            total_score += 0.05
        else:
            print(f"FAIL: Component 8 — Number formats ({correct_formats}/4 columns correct)")
    except Exception as e:
        print(f"ERROR: Component 8 — {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
