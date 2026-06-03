"""
Initial Setup: Raw energy usage data for formatting task
Task ID: calc_gpm_081
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gpm_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Energy"

    # Row 1: Simple title (unmerged, unformatted)
    ws["A1"] = "Home Energy Usage Analysis - 2025"

    # Row 3: Plain headers (no formatting)
    headers = ["Month", "kWh Used", "Rate ($/kWh)", "Electric Bill",
               "Gas Therms", "Gas Bill", "Total Bill"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)

    # Rows 4-15: 12 months of raw data with realistic seasonal patterns
    # kWh: high in summer (Jun-Aug) and winter (Dec-Feb), moderate spring/fall
    # Gas therms: high in winter (Nov-Mar), low in summer
    months_data = [
        # Month, kWh, Rate($/kWh), ElecBill, GasTherms, GasBill, TotalBill
        ["January",   680, None, None, 145, None, None],
        ["February",  620, None, None, 135, None, None],
        ["March",     480, None, None, 105, None, None],
        ["April",     390, None, None,  65, None, None],
        ["May",       420, None, None,  35, None, None],
        ["June",      750, None, None,  18, None, None],
        ["July",      880, None, None,  12, None, None],
        ["August",    845, None, None,  15, None, None],
        ["September", 560, None, None,  28, None, None],
        ["October",   430, None, None,  72, None, None],
        ["November",  550, None, None, 118, None, None],
        ["December",  720, None, None, 152, None, None],
    ]

    for r, row_data in enumerate(months_data, 4):
        ws.cell(row=r, column=1, value=row_data[0])  # Month
        ws.cell(row=r, column=2, value=row_data[1])  # kWh Used
        # Column C (Rate), D (Electric Bill), F (Gas Bill), G (Total Bill)
        # left as raw numbers - no formulas in initial
        ws.cell(row=r, column=5, value=row_data[4])  # Gas Therms

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
