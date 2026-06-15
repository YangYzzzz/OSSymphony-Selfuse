"""
Reward Script: Build a what-if data table showing how changes in both
average deal size and win rate affect quarterly revenue.
Task ID: calc_sales_081
Domain: libreoffice_calc
Scoring:
  Component 1 (0.20): B4 contains Expected Revenue formula =B1*B2*B3
  Component 2 (0.50): Data table body B8:G12 all contain formulas (30 cells)
  Component 3 (0.30): Formula correctness — formulas reference correct row/col cells
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_081'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'WhatIf' sheet must exist
    if 'WhatIf' not in wb.sheetnames:
        print(f"FAIL: Sheet 'WhatIf' not found. Sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['WhatIf']

    # ----------------------------------------------------------------
    # Component 1: B4 contains Expected Revenue formula (0.20 points)
    # Task requires B4 = B1*B2*B3. This cell is empty in initial_env.
    # ----------------------------------------------------------------
    try:
        b4_val = ws['B4'].value
        if b4_val is not None and isinstance(b4_val, str) and b4_val.startswith('='):
            # Check that the formula references B1, B2, B3 with multiplication
            formula_upper = b4_val.upper().replace(' ', '')
            refs_b1 = 'B1' in formula_upper
            refs_b2 = 'B2' in formula_upper
            refs_b3 = 'B3' in formula_upper
            has_mult = '*' in formula_upper
            if refs_b1 and refs_b2 and refs_b3 and has_mult:
                print(f"PASS: Component 1 — B4 has correct formula: {b4_val} (0.20 pts)")
                total_score += 0.20
            else:
                print(f"FAIL: Component 1 — B4 formula '{b4_val}' does not reference B1*B2*B3")
        else:
            print(f"FAIL: Component 1 — B4 does not contain a formula, found: {repr(b4_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # ----------------------------------------------------------------
    # Component 2: Data table body B8:G12 all contain formulas (0.50 pts)
    # In initial_env, these cells are all empty. In golden_env, each
    # of the 30 cells should contain a formula.
    # Score proportionally: 0.50 * (cells_with_formulas / 30)
    # ----------------------------------------------------------------
    try:
        formula_count = 0
        total_cells = 30  # 5 rows (8-12) x 6 cols (B-G)
        for row in range(8, 13):
            for col in range(2, 8):  # B=2, G=7
                cell_val = ws.cell(row=row, column=col).value
                if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                    formula_count += 1

        if formula_count == total_cells:
            print(f"PASS: Component 2 — All {total_cells} data table cells contain formulas (0.50 pts)")
            total_score += 0.50
        elif formula_count > 0:
            partial = round(0.50 * (formula_count / total_cells), 4)
            print(f"PARTIAL: Component 2 — {formula_count}/{total_cells} cells have formulas ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 2 — No formulas found in B8:G12 data table body")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # ----------------------------------------------------------------
    # Component 3: Formula correctness — formulas reference correct
    # deal size rows and win rate columns (0.30 pts)
    #
    # Expected pattern for columns C-G:
    #   =50*$<col>$6*$A$<row>  (references win rate header and deal size)
    # For column B:
    #   =50*$B$2*$A$<row>  (references base win rate input cell)
    #
    # We spot-check 6 representative cells.
    # ----------------------------------------------------------------
    try:
        correct_count = 0
        checks_total = 6

        # Define spot checks: (row, col, expected_pattern_fragments)
        # Each check verifies the formula references the right row and column
        spot_checks = [
            # (row, col, description, must_contain_list)
            (8, 3, "C8: =50*$C$6*$A$8", ['50', 'C$6', 'A$8']),
            (8, 5, "E8: =50*$E$6*$A$8", ['50', 'E$6', 'A$8']),
            (10, 4, "D10: =50*$D$6*$A$10", ['50', 'D$6', 'A$10']),
            (10, 7, "G10: =50*$G$6*$A$10", ['50', 'G$6', 'A$10']),
            (12, 7, "G12: =50*$G$6*$A$12", ['50', 'G$6', 'A$12']),
            (8, 2, "B8: =50*$B$2*$A$8", ['50', 'B$2', 'A$8']),
        ]

        for row, col, desc, fragments in spot_checks:
            cell_val = ws.cell(row=row, column=col).value
            if cell_val is not None and isinstance(cell_val, str) and cell_val.startswith('='):
                formula_upper = cell_val.upper().replace(' ', '')
                if all(frag.upper() in formula_upper for frag in fragments):
                    correct_count += 1
                    print(f"  CHECK OK: {desc} -> {cell_val}")
                else:
                    print(f"  CHECK FAIL: {desc} -> got {cell_val}, expected fragments {fragments}")
            else:
                print(f"  CHECK FAIL: {desc} -> no formula, got {repr(cell_val)}")

        if correct_count == checks_total:
            print(f"PASS: Component 3 — All {checks_total} spot checks passed (0.30 pts)")
            total_score += 0.30
        elif correct_count > 0:
            partial = round(0.30 * (correct_count / checks_total), 4)
            print(f"PARTIAL: Component 3 — {correct_count}/{checks_total} spot checks passed ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 3 — No spot checks passed")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(round(total_score, 2), 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
