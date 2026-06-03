"""
Initial Setup: Build a what-if data table showing how changes in both average deal size and win rate affect quarterly revenue.
Task ID: calc_sales_081
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_sales_081'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = 'WhatIf'

    # --- Parameters section ---
    ws['A1'] = 'Quarterly Opportunities'
    ws['B1'] = 50

    ws['A2'] = 'Win Rate'
    ws['B2'] = 0.30

    ws['A3'] = 'Avg Deal Size'
    ws['B3'] = 50000

    ws['A4'] = 'Expected Revenue'
    # B4 is intentionally EMPTY - the task is to create the formula and data table

    # --- Data Table header ---
    ws['A6'] = 'Data Table: Revenue by Win Rate and Deal Size'

    ws['B6'] = 'Win Rate >'
    ws['C6'] = 0.20
    ws['D6'] = 0.25
    ws['E6'] = 0.30
    ws['F6'] = 0.35
    ws['G6'] = 0.40

    ws['A7'] = 'Deal Size v'

    # Deal size row headers
    ws['A8'] = 30000
    ws['A9'] = 40000
    ws['A10'] = 50000
    ws['A11'] = 60000
    ws['A12'] = 70000

    # B8:G12 intentionally EMPTY - the task is to fill these with formulas

    # --- Formatting for readability ---
    # Bold parameter labels
    header_font = Font(bold=True)
    for cell_ref in ['A1', 'A2', 'A3', 'A4', 'A6', 'B6', 'A7']:
        ws[cell_ref].font = header_font

    # Format win rates as percentages
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws[f'{col_letter}6'].number_format = '0%'

    # Format B2 as percentage
    ws['B2'].number_format = '0%'

    # Format deal sizes and B3 as currency
    ws['B3'].number_format = '$#,##0'
    for row in range(8, 13):
        ws.cell(row=row, column=1).number_format = '$#,##0'

    # Adjust column widths
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 14
    for col_letter in ['C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col_letter].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
