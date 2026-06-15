"""
FINAL REWARD SCRIPT - SUCCESS
Task: I want to copy 'Unformatted Song Titles' to 'Clean Song Titles'. Please eliminate extra spaces and use proper case formatting (first letter capitalized per word, rest lowercase). Complete this without modifying blank cells.
Generated: 2025-11-24 07:32:25
Status: success
Model: o3
Total Steps: 1
"""

import openpyxl
import re
import os


def normalize_title(title: str) -> str:
    """Collapse extra spaces and convert to Proper Case."""
    collapsed = re.sub(r"\s+", " ", title.strip())  # remove leading/trailing & duplicate spaces
    return collapsed.title()  # each word capitalized, rest lowercase


def verify_song_titles(file_path: str) -> float:
    """Verify that the task of cleaning song titles was completed correctly.

    Scoring:
    - Each non-header data row is worth equal fractional points.
    - A row is correct if:
        * Both original & clean cells are blank  → counts as correct (blank preserved)
        * Original is non-blank and Clean equals normalize_title(original)
        * Original is blank and Clean is also blank (no unwanted text added)
    - Final score = correct_rows / total_rows  (progressive up to 1.0)
    """
    print(f"Verifying file: {file_path}")

    if not os.path.isfile(file_path):
        print("✗ File not found")
        return 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"✗ Unable to load workbook: {e}")
        return 0.0

    if 'Song List' not in wb.sheetnames:
        print("✗ Required sheet 'Song List' is missing")
        return 0.0

    sheet = wb['Song List']

    # Validate headers (no points, but must match to continue)
    if str(sheet['A1'].value).strip() != 'Unformatted Song Titles' or \
       str(sheet['B1'].value).strip() != 'Clean Song Titles':
        print("✗ Headers are incorrect. Expected 'Unformatted Song Titles' and 'Clean Song Titles'.")
        return 0.0

    total_rows = 0
    correct_rows = 0

    # Iterate over data rows (starting at row 2)
    for row in range(2, sheet.max_row + 1):
        orig_val = sheet.cell(row=row, column=1).value
        clean_val = sheet.cell(row=row, column=2).value

        orig_blank = orig_val is None or str(orig_val).strip() == ''
        clean_blank = clean_val is None or str(clean_val).strip() == ''

        if orig_blank and clean_blank:
            correct = True  # blank preserved
        elif orig_blank and not clean_blank:
            correct = False  # should not add text where none existed
        else:
            expected_clean = normalize_title(str(orig_val))
            correct = clean_val == expected_clean
            if not correct:
                print(f"✗ Row {row} mismatch -> Original: '{orig_val}', Expected Clean: '{expected_clean}', Found Clean: '{clean_val}'")

        total_rows += 1
        if correct:
            correct_rows += 1

    if total_rows == 0:
        print("✗ No data rows found to evaluate")
        return 0.0

    score = correct_rows / total_rows
    score = round(min(max(score, 0.0), 1.0), 4)  # clamp & round

    print(f"Correct rows: {correct_rows}/{total_rows}")
    print(f"REWARD: {score}")
    return score


# -----------------------------------------------------------------------
# Execute verification when the script runs directly
if __name__ == "__main__":
    FILE_PATH = "/home/user/i_want_to_copy_unformatted_song_titles_to_clean_song_titles_please_eliminate_extra_spaces_and_use_pr.xlsx"
    verify_song_titles(FILE_PATH)
