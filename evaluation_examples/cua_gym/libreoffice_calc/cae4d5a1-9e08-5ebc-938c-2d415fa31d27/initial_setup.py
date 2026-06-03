"""
Initial Setup: Create spreadsheet with print area limited to A1:D50
Task ID: calc_adv_print_clear_019
Domain: libreoffice_calc
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_adv_print_clear_019'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: Full Report ---
    ws = wb.active
    ws.title = 'Full Report'

    # Headers (row 1) across 8 columns
    headers = [
        'Employee ID', 'Full Name', 'Department', 'Job Title',
        'Base Salary', 'Bonus', 'Total Compensation', 'Hire Date'
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF', size=11)

    # Department list for realistic data
    departments = [
        'Engineering', 'Marketing', 'Sales', 'Finance', 'HR',
        'Operations', 'Product', 'Legal', 'Design', 'Support'
    ]
    titles_by_dept = {
        'Engineering': ['Software Engineer', 'Senior Engineer', 'Tech Lead', 'Principal Engineer'],
        'Marketing': ['Marketing Manager', 'Content Strategist', 'Brand Analyst', 'Growth Lead'],
        'Sales': ['Account Executive', 'Sales Manager', 'Business Dev Rep', 'Regional Director'],
        'Finance': ['Financial Analyst', 'Senior Accountant', 'Controller', 'Finance Manager'],
        'HR': ['HR Specialist', 'Recruiter', 'HR Business Partner', 'Talent Manager'],
        'Operations': ['Operations Analyst', 'Process Manager', 'Supply Chain Lead', 'Ops Director'],
        'Product': ['Product Manager', 'Senior PM', 'Product Analyst', 'VP of Product'],
        'Legal': ['Legal Counsel', 'Compliance Officer', 'Contract Manager', 'General Counsel'],
        'Design': ['UX Designer', 'UI Designer', 'Design Lead', 'Creative Director'],
        'Support': ['Support Specialist', 'Customer Success', 'Support Lead', 'CS Manager'],
    }

    employee_names = [
        'Sarah Chen', 'Marcus Johnson', 'Emily Rodriguez', 'David Kim',
        'Priya Sharma', 'James O\'Brien', 'Aisha Patel', 'Carlos Mendez',
        'Natalie Webb', 'Thomas Laurent', 'Lisa Nguyen', 'Robert Fischer',
        'Monica Castillo', 'Andrew Park', 'Fatima Al-Hassan', 'Steven Kowalski',
        'Yuki Tanaka', 'Grace Okafor', 'Daniel Brennan', 'Zoe Andersen',
        'Michael Torres', 'Rachel Goldstein', 'Kevin Wu', 'Amanda Foster',
        'Brian Nakamura', 'Sophia Delacroix', 'Patrick Murphy', 'Ingrid Johansson',
        'Omar Abdullah', 'Jennifer Blackwell', 'Christopher Lim', 'Valentina Cruz',
        'Sean Fitzgerald', 'Mei-Ling Huang', 'Aaron Okonkwo', 'Isabelle Moreau',
        'Tyler Henderson', 'Ananya Iyer', 'Caleb Simmons', 'Diana Petrov',
        'Nathan Carey', 'Leila Hosseini', 'Jonathan Adeyemi', 'Claire Dubois',
        'Brandon Reyes', 'Yolanda Baptiste', 'Lucas Schneider', 'Alicia Vang',
        'Derrick Powell', 'Hana Suzuki', 'Vincent Marchand', 'Tasha Williams',
        'Elliot Drummond', 'Nadia Korolev', 'Felix Kamau', 'Brooke Hartley',
        'Rodrigo Vasquez', 'Mei Chen', 'Declan Walsh', 'Amara Diallo',
        'Harrison Cole', 'Sunita Rao', 'Wesley Achebe', 'Camille Perrot',
        'Ian Mackenzie', 'Roya Dehghani', 'Dominic Esposito', 'Patience Nwachukwu',
        'Edward Bergstrom', 'Layla Mansour', 'Garrison Trent', 'Freya Lindqvist',
        'Malcolm Osei', 'Saoirse Quinn', 'Alberto Ferrante', 'Keiko Yamamoto',
        'Tobias Gruber', 'Miriam Abramowitz', 'Kwame Asante', 'Lucia Romano',
        'Shane Gallagher', 'Yuki Watanabe', 'Ibrahim Bakr', 'Stella Voronova',
        'Conrad Brandt', 'Adaeze Eze', 'Pierre Fontaine', 'Ximena Gutierrez',
        'Desmond Achterberg', 'Zainab Farooq', 'Florian Richter', 'Temi Adeyefa',
        'Hugo Leblanc', 'Sasha Volkov', 'Kwabena Mensah', 'Beatrice Johansson',
        'Rashid Al-Farsi', 'Chiara Bianchi', 'Oluwaseun Adegoke', 'Marie Cloutier',
        'Ethan Blackwood', 'Nilufar Yusupova', 'Gideon Osei', 'Theodora Stavros',
        'Femi Abiodun', 'Colette Renard', 'Arjun Nair', 'Brigitte Hoffmann',
        'Seun Oladipo', 'Helene Christensen', 'Babatunde Adewale', 'Fleur Moreau',
        'Chukwuemeka Obi', 'Astrid Eriksson', 'Seyi Afolabi', 'Mirko Petrovic',
        'Yewande Okonkwo', 'Lars Andersen', 'Adewale Abimbola', 'Sigrid Nielsen',
    ]

    import random
    random.seed(42)

    for i, name in enumerate(employee_names[:120], 2):
        dept_idx = (i - 2) % len(departments)
        dept = departments[dept_idx]
        title_list = titles_by_dept[dept]
        title = title_list[(i - 2) // len(departments) % len(title_list)]

        emp_id = f'EMP{1000 + i - 2:04d}'
        base_salary = random.randint(55, 145) * 1000
        bonus = round(base_salary * random.uniform(0.05, 0.20))
        total_comp = base_salary + bonus
        hire_year = random.randint(2015, 2024)
        hire_month = random.randint(1, 12)
        hire_day = random.randint(1, 28)
        hire_date = f'{hire_year}-{hire_month:02d}-{hire_day:02d}'

        ws.cell(row=i, column=1, value=emp_id)
        ws.cell(row=i, column=2, value=name)
        ws.cell(row=i, column=3, value=dept)
        ws.cell(row=i, column=4, value=title)
        ws.cell(row=i, column=5, value=base_salary)
        ws.cell(row=i, column=6, value=bonus)
        ws.cell(row=i, column=7, value=total_comp)
        ws.cell(row=i, column=8, value=hire_date)

    # Set column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 14

    # Freeze header row
    ws.freeze_panes = 'A2'

    # CRITICAL: Set print area to A1:D50 (this is the task condition that needs to be cleared)
    ws.print_area = 'A1:D50'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'  - Sheet: Full Report')
    print(f'  - Data range: A1:H120 (120 rows, 8 columns)')
    print(f'  - Print area: A1:D50 (limited, cutting off columns E-H and rows 51-120)')


create_initial()
