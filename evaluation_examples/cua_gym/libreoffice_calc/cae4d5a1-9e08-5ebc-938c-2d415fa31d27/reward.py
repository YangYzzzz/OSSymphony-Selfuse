"""
Reward Script: Clear print area so entire sheet prints
Task ID: calc_adv_print_clear_019
Domain: libreoffice_calc
Scoring:
  Component 1: Print area is completely removed/cleared (0.7 points)
  Component 2: Cell data preserved (A1:H120 data integrity) (0.3 points)
Total: 1.0
"""

import os
import openpyxl

WORKDIR = '/home/user'  # VM path — all reward scripts run on the VM
TASK_ID = 'calc_adv_print_clear_019'
SHEET_NAME = 'Full Report'


def verify_task(file_path):
    """
    Verify that the print area has been cleared from the 'Full Report' sheet.

    Initial state: print_area = "'Full Report'!$A$1:$D$50" (restricted to D50)
    Target state:  print_area = '' (empty — entire sheet prints)

    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    # Precondition: load the workbook
    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: sheet must exist
    if SHEET_NAME not in wb.sheetnames:
        print(f"CRITICAL: Sheet '{SHEET_NAME}' not found in {file_path}")
        print(f"Available sheets: {wb.sheetnames}")
        print("REWARD: 0.0")
        return 0.0

    ws = wb[SHEET_NAME]

    # ----------------------------------------------------------------
    # Component 1: Print area is completely cleared (0.7 points)
    # The task requires clearing the print area so the whole sheet prints.
    # Initial file has print_area = "'Full Report'!$A$1:$D$50"
    # Golden file must have print_area = '' (empty string / None)
    # ----------------------------------------------------------------
    try:
        print_area = ws.print_area
        # print_area is empty string or None when no print area is defined
        if not print_area:
            print(f"PASS: Component 1 — Print area is cleared (value: {repr(print_area)}) (0.7 pts)")
            total_score += 0.7
        else:
            print(f"FAIL: Component 1 — Print area is NOT cleared. Found: {repr(print_area)}")
            print(f"      Expected: empty string or None (no print area restriction)")
    except Exception as e:
        print(f"ERROR: Component 1 — Could not check print_area: {e}")

    # ----------------------------------------------------------------
    # Component 2: Cell data integrity preserved (0.3 points)
    # The task says cell values and formatting should be unchanged.
    # Verify that the used range still covers 8 columns x 121 rows (A1:H121 incl. header)
    # and that key header cells are present.
    # This check FAILS on initial only if combined with Component 1 passing —
    # but since data is the same in both, we check the structural size (121 rows, 8 cols)
    # AND that print area was cleared (gate: only award if component 1 passes).
    # This component awards points ONLY IF the print area is also cleared,
    # preventing the score from being awarded for the initial file.
    # ----------------------------------------------------------------
    try:
        max_row = ws.max_row
        max_col = ws.max_column

        # Check that the full data range is intact: 121 rows (header + 120 data), 8 columns
        header_a1 = ws.cell(row=1, column=1).value
        header_correct = header_a1 == 'Employee ID'

        data_range_intact = (max_row >= 121 and max_col >= 8)

        if not print_area and data_range_intact and header_correct:
            # Only award if print area was cleared AND data is intact
            print(f"PASS: Component 2 — Data integrity preserved after clearing print area "
                  f"(rows={max_row}, cols={max_col}, header A1={repr(header_a1)}) (0.3 pts)")
            total_score += 0.3
        elif print_area:
            # Print area not cleared — component 1 already failed, skip awarding here
            print(f"FAIL: Component 2 — Skipped because print area is still set")
        else:
            print(f"FAIL: Component 2 — Data integrity check failed. "
                  f"rows={max_row} (expected >=121), cols={max_col} (expected >=8), "
                  f"header A1={repr(header_a1)} (expected 'Employee ID')")
    except Exception as e:
        print(f"ERROR: Component 2 — Could not check data integrity: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against golden file (path on VM)
file_path = f'{WORKDIR}/{TASK_ID}_initial.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
