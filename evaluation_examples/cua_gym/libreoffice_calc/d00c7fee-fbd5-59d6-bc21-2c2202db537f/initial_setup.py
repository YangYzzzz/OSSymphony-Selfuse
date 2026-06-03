"""
Initial Setup: Conditional number formatting task - pre-task state
Task ID: calc_lf_087
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_087'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet: KPIs ---
    ws = wb.active
    ws.title = 'KPIs'

    # Header row
    headers = ['KPI', 'Score']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")

    # Data rows - KPI scores (NO custom number format applied - that is the task)
    data = [
        ['Customer Sat', 92],
        ['NPS', 115],
        ['Retention', 88],
        ['Referrals', 105],
    ]
    for r, (kpi, score) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=kpi).font = Font(name='Calibri', size=11)
        cell_b = ws.cell(row=r, column=2, value=score)
        cell_b.font = Font(name='Calibri', size=11)
        cell_b.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 12

    # --- Sheet 2: Targets (extra context for realism) ---
    ws2 = wb.create_sheet('Targets')
    ws2.cell(row=1, column=1, value='KPI').font = Font(bold=True)
    ws2.cell(row=1, column=2, value='Target').font = Font(bold=True)
    ws2.cell(row=1, column=3, value='Quarter').font = Font(bold=True)
    targets_data = [
        ['Customer Sat', 95, 'Q1 2025'],
        ['NPS', 110, 'Q1 2025'],
        ['Retention', 90, 'Q1 2025'],
        ['Referrals', 100, 'Q1 2025'],
        ['Customer Sat', 97, 'Q2 2025'],
        ['NPS', 112, 'Q2 2025'],
        ['Retention', 92, 'Q2 2025'],
        ['Referrals', 102, 'Q2 2025'],
        ['Customer Sat', 93, 'Q3 2025'],
        ['NPS', 108, 'Q3 2025'],
        ['Retention', 87, 'Q3 2025'],
        ['Referrals', 98, 'Q3 2025'],
    ]
    for r, row_data in enumerate(targets_data, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)
    ws2.column_dimensions['A'].width = 18
    ws2.column_dimensions['B'].width = 10
    ws2.column_dimensions['C'].width = 12

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
