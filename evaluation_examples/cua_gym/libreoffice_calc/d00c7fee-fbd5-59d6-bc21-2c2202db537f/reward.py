"""
Reward Script: Apply conditional number formatting to KPI scores
Task ID: calc_lf_087
Domain: libreoffice_calc
Scoring:
  Component 1: B2 number format is '[GREEN][>100]#,##0;[RED]#,##0' (0.25 pts)
  Component 2: B3 number format is '[GREEN][>100]#,##0;[RED]#,##0' (0.25 pts)
  Component 3: B4 number format is '[GREEN][>100]#,##0;[RED]#,##0' (0.25 pts)
  Component 4: B5 number format is '[GREEN][>100]#,##0;[RED]#,##0' (0.25 pts)
"""

import os
import time

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_087'
EXPECTED_FORMAT = '[GREEN][>100]#,##0;[RED]#,##0'


def persist_app_state(domain: str):
    """Attempt to save any unsaved LibreOffice work before verification."""
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify that cells B2:B5 on the 'KPIs' sheet have the conditional
    number format '[GREEN][>100]#,##0;[RED]#,##0' applied.
    Returns: float between 0.0 and 1.0
    """
    import openpyxl

    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Verify KPIs sheet exists
    if 'KPIs' not in wb.sheetnames:
        print("FAIL: Sheet 'KPIs' not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['KPIs']

    # Check each cell in B2:B5 for the correct number format
    cells_to_check = {
        'B2': ('Customer Sat', 92),
        'B3': ('NPS', 115),
        'B4': ('Retention', 88),
        'B5': ('Referrals', 105),
    }

    for cell_ref, (kpi_name, expected_val) in cells_to_check.items():
        row_num = int(cell_ref[1])
        component_num = row_num - 1  # Component 1-4

        # Component N: Cell has the conditional number format (0.25 points)
        try:
            cell = ws[cell_ref]
            actual_format = cell.number_format

            if actual_format == EXPECTED_FORMAT:
                print(f"PASS: Component {component_num} -- {cell_ref} ({kpi_name}) "
                      f"has correct format: '{actual_format}' (0.25 pts)")
                total_score += 0.25
            else:
                print(f"FAIL: Component {component_num} -- {cell_ref} ({kpi_name}) "
                      f"expected format '{EXPECTED_FORMAT}', found '{actual_format}'")
        except Exception as e:
            print(f"ERROR: Component {component_num} -- Could not check {cell_ref}: {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
