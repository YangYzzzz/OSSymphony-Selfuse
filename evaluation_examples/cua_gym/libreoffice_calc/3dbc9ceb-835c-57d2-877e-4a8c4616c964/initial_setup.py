"""
Initial Setup: Create spreadsheet with categories and amounts for pivot summary task
Task ID: calc_mcp_032
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_032'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    # Headers
    ws.cell(row=1, column=1, value='Category')
    ws.cell(row=1, column=2, value='Amount')

    # Make headers bold
    from openpyxl.styles import Font
    bold_font = Font(bold=True)
    ws['A1'].font = bold_font
    ws['B1'].font = bold_font

    # Realistic repeated categories with varied amounts (49 data rows: A2:A50, B2:B50)
    categories = ['Food', 'Transport', 'Rent', 'Entertainment', 'Utilities',
                  'Healthcare', 'Clothing', 'Education']

    data = [
        ('Food', 45.30),
        ('Transport', 32.00),
        ('Rent', 1200.00),
        ('Entertainment', 18.50),
        ('Utilities', 85.00),
        ('Healthcare', 120.00),
        ('Clothing', 65.00),
        ('Education', 250.00),
        ('Food', 62.75),
        ('Transport', 15.50),
        ('Food', 38.20),
        ('Entertainment', 45.00),
        ('Utilities', 92.30),
        ('Rent', 1200.00),
        ('Transport', 28.00),
        ('Healthcare', 55.00),
        ('Food', 29.90),
        ('Clothing', 89.50),
        ('Education', 175.00),
        ('Food', 51.40),
        ('Transport', 42.00),
        ('Entertainment', 32.00),
        ('Utilities', 78.50),
        ('Healthcare', 200.00),
        ('Food', 27.60),
        ('Rent', 1200.00),
        ('Transport', 18.75),
        ('Clothing', 120.00),
        ('Education', 99.00),
        ('Food', 43.80),
        ('Entertainment', 55.00),
        ('Utilities', 105.20),
        ('Healthcare', 75.00),
        ('Transport', 35.50),
        ('Food', 58.15),
        ('Rent', 1200.00),
        ('Clothing', 45.00),
        ('Education', 310.00),
        ('Food', 33.25),
        ('Transport', 22.00),
        ('Entertainment', 28.50),
        ('Utilities', 88.00),
        ('Healthcare', 150.00),
        ('Food', 41.60),
        ('Clothing', 72.50),
        ('Transport', 19.00),
        ('Education', 425.00),
        ('Food', 36.50),
        ('Rent', 1200.00),
    ]

    for r, (cat, amt) in enumerate(data, 2):
        ws.cell(row=r, column=1, value=cat)
        ws.cell(row=r, column=2, value=amt)

    # Set column widths for readability
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14

    # Number format for amounts
    for r in range(2, 51):
        ws.cell(row=r, column=2).number_format = '#,##0.00'

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
