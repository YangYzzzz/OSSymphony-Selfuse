"""
Reward Script: PivotSummary macro - pivot categories with summed values
Task ID: calc_mcp_032
Domain: libreoffice_calc
Scoring:
  Component 1: PivotResult sheet exists (0.2)
  Component 2: Correct headers in PivotResult (0.1)
  Component 3: Correct number of unique categories (0.2)
  Component 4: All category names present (0.2)
  Component 5: All summed values correct (0.3)
"""

import os
from collections import defaultdict

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_mcp_032'

# Expected pivot results computed from Sheet1 source data
EXPECTED_CATEGORIES = {
    'Food': 468.45,
    'Transport': 212.75,
    'Rent': 6000.0,
    'Entertainment': 179.0,
    'Utilities': 449.0,
    'Healthcare': 600.0,
    'Clothing': 392.0,
    'Education': 1259.0,
}


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Component 1: PivotResult sheet exists (0.2 points)
    try:
        if 'PivotResult' in wb.sheetnames:
            print(f"PASS: Component 1 -- PivotResult sheet exists (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 1 -- PivotResult sheet not found. Sheets: {wb.sheetnames}")
            # No PivotResult sheet means all subsequent checks fail too
            final_score = min(total_score, 1.0)
            print(f"\nScore: {total_score}/1.0")
            print(f"REWARD: {final_score}")
            return final_score
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")
        final_score = min(total_score, 1.0)
        print(f"\nScore: {total_score}/1.0")
        print(f"REWARD: {final_score}")
        return final_score

    ws = wb['PivotResult']

    # Component 2: Correct headers (Category, Amount) in row 1 (0.1 points)
    try:
        header_a = ws.cell(row=1, column=1).value
        header_b = ws.cell(row=1, column=2).value
        header_a_ok = header_a is not None and str(header_a).strip().lower() == 'category'
        header_b_ok = header_b is not None and str(header_b).strip().lower() == 'amount'
        if header_a_ok and header_b_ok:
            print(f"PASS: Component 2 -- Headers correct: A1='{header_a}', B1='{header_b}' (0.1 pts)")
            total_score += 0.1
        else:
            print(f"FAIL: Component 2 -- Expected headers 'Category','Amount', found '{header_a}','{header_b}'")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Read all data from PivotResult (rows 2+)
    pivot_data = {}
    try:
        for r in range(2, ws.max_row + 1):
            cat = ws.cell(row=r, column=1).value
            amt = ws.cell(row=r, column=2).value
            if cat is not None:
                cat_str = str(cat).strip()
                try:
                    amt_val = float(amt) if amt is not None else None
                except (ValueError, TypeError):
                    amt_val = None
                pivot_data[cat_str] = amt_val
    except Exception as e:
        print(f"ERROR: Reading PivotResult data -- {e}")

    print(f"  Found {len(pivot_data)} categories in PivotResult: {list(pivot_data.keys())}")

    # Component 3: Correct number of unique categories (0.2 points)
    try:
        expected_count = len(EXPECTED_CATEGORIES)  # 8
        actual_count = len(pivot_data)
        if actual_count == expected_count:
            print(f"PASS: Component 3 -- Correct category count: {actual_count} (0.2 pts)")
            total_score += 0.2
        else:
            print(f"FAIL: Component 3 -- Expected {expected_count} categories, found {actual_count}")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: All category names present (0.2 points)
    try:
        expected_names = set(EXPECTED_CATEGORIES.keys())
        actual_names = set(pivot_data.keys())
        if expected_names == actual_names:
            print(f"PASS: Component 4 -- All category names match (0.2 pts)")
            total_score += 0.2
        else:
            missing = expected_names - actual_names
            extra = actual_names - expected_names
            print(f"FAIL: Component 4 -- Missing: {missing}, Extra: {extra}")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    # Component 5: All summed values correct within tolerance (0.3 points)
    try:
        tolerance = 0.01
        correct_count = 0
        total_expected = len(EXPECTED_CATEGORIES)
        for cat, expected_val in EXPECTED_CATEGORIES.items():
            actual_val = pivot_data.get(cat)
            if actual_val is not None and abs(actual_val - expected_val) <= tolerance:
                correct_count += 1
            else:
                print(f"  Value mismatch for '{cat}': expected {expected_val}, got {actual_val}")

        if correct_count == total_expected:
            print(f"PASS: Component 5 -- All {total_expected} summed values correct (0.3 pts)")
            total_score += 0.3
        elif correct_count > 0:
            partial = round(0.3 * correct_count / total_expected, 2)
            print(f"PARTIAL: Component 5 -- {correct_count}/{total_expected} values correct ({partial} pts)")
            total_score += partial
        else:
            print(f"FAIL: Component 5 -- No summed values match")
    except Exception as e:
        print(f"ERROR: Component 5 -- {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Default: test against canonical artifact path
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
