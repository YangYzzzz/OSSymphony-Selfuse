"""
Reward Script: Apply bold, italic, and red font color to cells C2:C10
Task ID: calc_gfl_091
Domain: libreoffice_calc
Scoring:
  Component 1 (0.35): All cells C2:C10 are bold
  Component 2 (0.35): All cells C2:C10 are italic
  Component 3 (0.20): All cells C2:C10 have red font color
  Component 4 (0.10): Cells C11:C20 retain original formatting (not bold, not italic)
"""

import os

import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_091'


def persist_app_state(domain: str):
    """Save any unsaved GUI state before verification."""
    import time
    os.environ["DISPLAY"] = ":0"
    if domain in {"libreoffice_calc", "libreoffice_writer", "libreoffice_impress"}:
        try:
            import pyautogui
            pyautogui.hotkey("ctrl", "s")
            time.sleep(0.8)
            print(f"PERSIST: ctrl+s sent for {domain}")
        except Exception as e:
            print(f"PERSIST_WARN: save hook failed: {e}")


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: 'QA' sheet must exist
    if 'QA' not in wb.sheetnames:
        print("CRITICAL: 'QA' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['QA']

    # Component 1: All cells C2:C10 are bold (0.35 points)
    try:
        bold_count = 0
        for r in range(2, 11):
            cell = ws.cell(row=r, column=3)
            if cell.font.bold is True:
                bold_count += 1
        if bold_count == 9:
            print(f"PASS: Component 1 -- All 9 cells C2:C10 are bold (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 1 -- Only {bold_count}/9 cells in C2:C10 are bold")
    except Exception as e:
        print(f"ERROR: Component 1 -- {e}")

    # Component 2: All cells C2:C10 are italic (0.35 points)
    try:
        italic_count = 0
        for r in range(2, 11):
            cell = ws.cell(row=r, column=3)
            if cell.font.italic is True:
                italic_count += 1
        if italic_count == 9:
            print(f"PASS: Component 2 -- All 9 cells C2:C10 are italic (0.35 pts)")
            total_score += 0.35
        else:
            print(f"FAIL: Component 2 -- Only {italic_count}/9 cells in C2:C10 are italic")
    except Exception as e:
        print(f"ERROR: Component 2 -- {e}")

    # Component 3: All cells C2:C10 have red font color (0.20 points)
    # Red in ARGB can be 00FF0000 or FFFF0000 (alpha varies)
    try:
        red_count = 0
        for r in range(2, 11):
            cell = ws.cell(row=r, column=3)
            try:
                color_rgb = cell.font.color.rgb if cell.font.color else None
                if color_rgb and isinstance(color_rgb, str) and color_rgb.upper().endswith('FF0000'):
                    red_count += 1
                else:
                    print(f"  C{r} font color: {color_rgb}")
            except (TypeError, AttributeError):
                print(f"  C{r} font color: theme/default (not red)")
        if red_count == 9:
            print(f"PASS: Component 3 -- All 9 cells C2:C10 have red font color (0.20 pts)")
            total_score += 0.20
        else:
            print(f"FAIL: Component 3 -- Only {red_count}/9 cells in C2:C10 have red font")
    except Exception as e:
        print(f"ERROR: Component 3 -- {e}")

    # Component 4: C2:C10 are bold+italic+red AND C11:C20 retain default formatting (0.10 points)
    # This is a compound check anchored to the task change -- only awards if the task was done
    # AND the passed-cases cells were left alone.
    try:
        # First check: at least some C2:C10 cells were formatted (task change happened)
        task_change_detected = (bold_count > 0 and italic_count > 0 and red_count > 0)
        if not task_change_detected:
            print(f"FAIL: Component 4 -- Task change not detected, skipping C11:C20 check")
        else:
            unchanged_count = 0
            for r in range(11, 21):
                cell = ws.cell(row=r, column=3)
                if cell.font.bold is not True and cell.font.italic is not True:
                    unchanged_count += 1
                else:
                    print(f"  C{r}: bold={cell.font.bold}, italic={cell.font.italic} (should be default)")
            if unchanged_count == 10:
                print(f"PASS: Component 4 -- Task applied AND C11:C20 retain original formatting (0.10 pts)")
                total_score += 0.10
            else:
                print(f"FAIL: Component 4 -- {10 - unchanged_count}/10 cells in C11:C20 were modified")
    except Exception as e:
        print(f"ERROR: Component 4 -- {e}")

    final_score = round(min(total_score, 1.0), 2)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
persist_app_state("libreoffice_calc")

file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
