"""
Initial Setup: Apply bold, italic, and red font color to error cells C2:C10
Task ID: calc_gfl_091
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_gfl_091'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'QA'

    # Row 1: Headers
    headers = ['Test Case', 'Expected', 'Result', 'Actual', 'Pass/Fail', 'Error Code']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Rows 2-10: Failed test cases (9 rows)
    failed_data = [
        ['TC-1001 Login Timeout',        200, 504, 'Gateway Timeout',     'Fail', 'ERR-5001'],
        ['TC-1002 API Auth Failure',      200, 401, 'Unauthorized',        'Fail', 'ERR-4001'],
        ['TC-1003 DB Connection Drop',    'OK', 'ERR', 'Connection reset', 'Fail', 'ERR-6010'],
        ['TC-1004 File Upload Limit',     10,  413, 'Payload Too Large',   'Fail', 'ERR-4130'],
        ['TC-1005 Session Expiry',        200, 403, 'Forbidden',           'Fail', 'ERR-4003'],
        ['TC-1006 Rate Limiter Block',    200, 429, 'Too Many Requests',   'Fail', 'ERR-4290'],
        ['TC-1007 SSL Cert Mismatch',     'OK', 'ERR', 'SSL handshake failed', 'Fail', 'ERR-7001'],
        ['TC-1008 Memory Overflow',       'OK', 'ERR', 'Out of memory',   'Fail', 'ERR-8001'],
        ['TC-1009 Null Pointer Exception', 200, 500, 'Internal Server Error', 'Fail', 'ERR-5000'],
    ]
    for r, row_data in enumerate(failed_data, 2):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Rows 11-20: Passed test cases (10 rows)
    passed_data = [
        ['TC-2001 User Registration',    200, 200, 'OK',   'Pass', ''],
        ['TC-2002 Password Reset',        200, 200, 'OK',   'Pass', ''],
        ['TC-2003 Profile Update',        200, 200, 'OK',   'Pass', ''],
        ['TC-2004 Search Query',          200, 200, 'OK',   'Pass', ''],
        ['TC-2005 Checkout Flow',         200, 200, 'OK',   'Pass', ''],
        ['TC-2006 Email Notification',    200, 200, 'OK',   'Pass', ''],
        ['TC-2007 Dashboard Load',        200, 200, 'OK',   'Pass', ''],
        ['TC-2008 Export CSV',            200, 200, 'OK',   'Pass', ''],
        ['TC-2009 Bulk Delete',           200, 200, 'OK',   'Pass', ''],
        ['TC-2010 Two-Factor Auth',       200, 200, 'OK',   'Pass', ''],
    ]
    for r, row_data in enumerate(passed_data, 11):
        for c, val in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=val)

    # Set reasonable column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 24
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
