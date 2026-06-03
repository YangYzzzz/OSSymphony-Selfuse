"""
Initial Setup: Conditional formatting on Sales sheet for profit analysis
Task ID: calc_ggf_044
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import random
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime, timedelta

WORKDIR = '/home/user'
TASK_ID = 'calc_ggf_044'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    random.seed(42)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    # --- Headers ---
    headers = ['Transaction ID', 'Date', 'Product', 'Revenue', 'Cost', 'Profit']
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Products list ---
    products = [
        'Wireless Headphones', 'USB-C Hub', 'Mechanical Keyboard', 'Webcam HD',
        'Monitor Stand', 'Laptop Sleeve', 'Mouse Pad XL', 'LED Desk Lamp',
        'Portable SSD 1TB', 'Bluetooth Speaker', 'Phone Charger', 'HDMI Cable',
        'Ergonomic Chair', 'Standing Desk Mat', 'Cable Management Kit',
        'Screen Protector', 'Tablet Stylus', 'Power Strip', 'Desk Organizer',
        'Noise Cancelling Mic'
    ]

    # --- Generate 50 rows of data (rows 2-51) ---
    base_date = datetime(2025, 1, 5)
    for i in range(50):
        row = i + 2
        # Transaction ID
        tid = f'TXN-2025-{1001 + i}'
        ws.cell(row=row, column=1, value=tid)

        # Date (spread across Jan-Dec 2025)
        date_val = base_date + timedelta(days=random.randint(0, 350))
        ws.cell(row=row, column=2, value=date_val)
        ws.cell(row=row, column=2).number_format = 'yyyy-mm-dd'

        # Product
        product = random.choice(products)
        ws.cell(row=row, column=3, value=product)

        # Revenue and Cost - designed to produce mix of negative, mid-range, and high profits
        if i % 7 == 0:
            # Negative profit cases (about 7-8 rows)
            revenue = round(random.uniform(200, 1500), 2)
            cost = round(revenue + random.uniform(50, 800), 2)
        elif i % 5 == 0:
            # High profit cases > 5000 (about 7-8 rows)
            revenue = round(random.uniform(8000, 18000), 2)
            cost = round(random.uniform(1500, 5000), 2)
        else:
            # Mid-range profit 0 <= profit <= 5000
            revenue = round(random.uniform(2000, 8000), 2)
            cost = round(random.uniform(1500, 6000), 2)
            # Clamp to ensure 0 <= profit <= 5000
            profit = revenue - cost
            if profit < 0:
                cost = revenue - random.uniform(100, 3000)
                cost = round(cost, 2)
            elif profit > 5000:
                cost = revenue - random.uniform(500, 4500)
                cost = round(cost, 2)

        ws.cell(row=row, column=4, value=revenue)
        ws.cell(row=row, column=4).number_format = '$#,##0.00'

        ws.cell(row=row, column=5, value=cost)
        ws.cell(row=row, column=5).number_format = '$#,##0.00'

        # Profit = Revenue - Cost (as formula)
        ws.cell(row=row, column=6, value=f'=D{row}-E{row}')
        ws.cell(row=row, column=6).number_format = '$#,##0.00'

    # --- Column widths ---
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14

    # Freeze header row
    ws.freeze_panes = "A2"

    # NO conditional formatting in initial state

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
