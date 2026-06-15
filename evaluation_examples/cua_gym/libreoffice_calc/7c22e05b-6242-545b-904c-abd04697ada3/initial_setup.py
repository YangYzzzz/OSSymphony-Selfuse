"""
Initial Setup: Lab Equipment Reservation Sheet (no data validation)
Task ID: calc_edu_equipment_reservation_036
Domain: libreoffice_calc
"""

import openpyxl
from datetime import date

WORKDIR = '/home/user'
TASK_ID = 'calc_edu_equipment_reservation_036'
OUTPUT = f'{WORKDIR}/{TASK_ID}_initial.xlsx'


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Reservations ---
    ws1 = wb.active
    ws1.title = 'Reservations'

    headers = ['Reservation ID', 'Equipment Name', 'Borrower', 'Checkout Date', 'Return Date', 'Status']
    for col, h in enumerate(headers, 1):
        ws1.cell(row=1, column=col, value=h)

    # 75 realistic reservation records
    equipment_names = [
        'Microscope', 'Centrifuge', 'Spectrophotometer', 'pH Meter',
        'Hot Plate Stirrer', 'Analytical Balance', 'Autoclave',
        'PCR Thermal Cycler', 'Gel Electrophoresis System', 'Oscilloscope',
        'Function Generator', 'Digital Multimeter', 'Soldering Station',
        '3D Printer', 'Laser Cutter'
    ]

    borrowers = [
        'Alice Wang', 'Ben Carter', 'Clara Kim', 'David Liu', 'Eva Patel',
        'Frank Zhang', 'Grace Nguyen', 'Henry Park', 'Isla Martinez', 'Jack Lee',
        'Karen Brooks', 'Liam Chen', 'Mia Thompson', 'Nathan Singh', 'Olivia Brown',
        'Peter Yao', 'Quinn Adams', 'Rachel Wu', 'Samuel Torres', 'Tina Zhao',
        'Uma Gupta', 'Victor Reyes', 'Wendy Huang', 'Xander Novak', 'Yara Hassan'
    ]

    statuses = ['Reserved', 'In Use', 'Returned']

    # Checkout dates in 2025
    checkout_dates = [
        '2025-01-05', '2025-01-12', '2025-01-20', '2025-01-27', '2025-02-03',
        '2025-02-10', '2025-02-18', '2025-02-24', '2025-03-04', '2025-03-11',
        '2025-03-18', '2025-03-25', '2025-04-01', '2025-04-08', '2025-04-15',
        '2025-04-22', '2025-04-29', '2025-05-06', '2025-05-13', '2025-05-20',
        '2025-05-27', '2025-06-03', '2025-06-10', '2025-06-17', '2025-06-24',
        '2025-07-01', '2025-07-08', '2025-07-15', '2025-07-22', '2025-07-29',
        '2025-08-05', '2025-08-12', '2025-08-19', '2025-08-26', '2025-09-02',
        '2025-09-09', '2025-09-16', '2025-09-23', '2025-09-30', '2025-10-07',
        '2025-10-14', '2025-10-21', '2025-10-28', '2025-11-04', '2025-11-11',
        '2025-11-18', '2025-11-25', '2025-12-02', '2025-12-09', '2025-12-16',
        '2025-01-08', '2025-02-14', '2025-03-07', '2025-04-03', '2025-05-01',
        '2025-06-06', '2025-07-04', '2025-08-08', '2025-09-05', '2025-10-03',
        '2025-01-15', '2025-02-21', '2025-03-14', '2025-04-11', '2025-05-09',
        '2025-06-13', '2025-07-11', '2025-08-15', '2025-09-12', '2025-10-10',
        '2025-01-22', '2025-02-28', '2025-03-21', '2025-04-18', '2025-05-16',
    ]

    return_dates = [
        '2025-01-12', '2025-01-19', '2025-01-27', '2025-02-03', '2025-02-10',
        '2025-02-17', '2025-02-25', '2025-03-03', '2025-03-11', '2025-03-18',
        '2025-03-25', '2025-04-01', '2025-04-08', '2025-04-15', '2025-04-22',
        '2025-04-29', '2025-05-06', '2025-05-13', '2025-05-20', '2025-05-27',
        '2025-06-03', '2025-06-10', '2025-06-17', '2025-06-24', '2025-07-01',
        '2025-07-08', '2025-07-15', '2025-07-22', '2025-07-29', '2025-08-05',
        '2025-08-12', '2025-08-19', '2025-08-26', '2025-09-02', '2025-09-09',
        '2025-09-16', '2025-09-23', '2025-09-30', '2025-10-07', '2025-10-14',
        '2025-10-21', '2025-10-28', '2025-11-04', '2025-11-11', '2025-11-18',
        '2025-11-25', '2025-12-02', '2025-12-09', '2025-12-16', '2025-12-23',
        '2025-01-15', '2025-02-21', '2025-03-14', '2025-04-10', '2025-05-08',
        '2025-06-13', '2025-07-11', '2025-08-15', '2025-09-12', '2025-10-10',
        '2025-01-22', '2025-02-28', '2025-03-21', '2025-04-18', '2025-05-16',
        '2025-06-20', '2025-07-18', '2025-08-22', '2025-09-19', '2025-10-17',
        '2025-01-29', '2025-03-07', '2025-03-28', '2025-04-25', '2025-05-23',
    ]

    for i in range(75):
        row = i + 2
        res_id = f'RES-2025-{i+1:03d}'
        eq_name = equipment_names[i % len(equipment_names)]
        borrower = borrowers[i % len(borrowers)]
        checkout = checkout_dates[i]
        ret = return_dates[i]
        status = statuses[i % len(statuses)]

        ws1.cell(row=row, column=1, value=res_id)
        ws1.cell(row=row, column=2, value=eq_name)
        ws1.cell(row=row, column=3, value=borrower)
        ws1.cell(row=row, column=4, value=checkout)
        ws1.cell(row=row, column=5, value=ret)
        ws1.cell(row=row, column=6, value=status)

    # --- Sheet 2: EquipmentList ---
    ws2 = wb.create_sheet('EquipmentList')

    approved_equipment = [
        'Microscope',
        'Centrifuge',
        'Spectrophotometer',
        'pH Meter',
        'Hot Plate Stirrer',
        'Analytical Balance',
        'Autoclave',
        'PCR Thermal Cycler',
        'Gel Electrophoresis System',
        'Oscilloscope',
        'Function Generator',
        'Digital Multimeter',
        'Soldering Station',
        '3D Printer',
        'Laser Cutter',
    ]

    for i, name in enumerate(approved_equipment, 1):
        ws2.cell(row=i, column=1, value=name)

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')
    print(f'Reservations sheet: 1 header row + 75 data rows')
    print(f'EquipmentList sheet: 15 approved equipment names in A1:A15')
    print('NOTE: No data validation rules set (task requires adding them)')


create_initial()
