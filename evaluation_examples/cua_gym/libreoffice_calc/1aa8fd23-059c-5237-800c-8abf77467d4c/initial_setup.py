"""
Initial Setup: Nested VLOOKUP - Category code to discount percentage lookup
Task ID: calc_lf_017
Domain: libreoffice_calc
"""

import os
import shlex
import subprocess
import time
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_017'
OUTPUT = f'{WORKDIR}/{TASK_ID}.xlsx'


def launch_gui(command: str, delay_sec: float = 1.0):
    """Launch GUI app on VM display without blocking script exit."""
    env = os.environ.copy()
    env["DISPLAY"] = ":0"
    subprocess.Popen(
        shlex.split(command),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        env=env,
    )
    time.sleep(delay_sec)


def create_initial():
    wb = openpyxl.Workbook()

    # --- Sheet 1: Products ---
    ws1 = wb.active
    ws1.title = 'Products'

    # Headers
    headers = ['Product', 'Category Code', 'Discount %']
    header_font = Font(name='Calibri', size=11, bold=True)
    header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
    header_font_white = Font(name='Calibri', size=11, bold=True, color='FFFFFF')

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # Product data
    products = [
        ['Laptop', 'CAT-A'],
        ['Mouse', 'CAT-C'],
        ['Monitor', 'CAT-B'],
    ]
    for r, row_data in enumerate(products, 2):
        for c, val in enumerate(row_data, 1):
            ws1.cell(row=r, column=c, value=val)

    # C2:C4 left EMPTY - task is to enter the nested VLOOKUP formula

    # Column widths
    ws1.column_dimensions['A'].width = 15
    ws1.column_dimensions['B'].width = 16
    ws1.column_dimensions['C'].width = 14

    # --- Sheet 2: Discounts ---
    ws2 = wb.create_sheet('Discounts')

    disc_headers = ['Category Code', 'Discount %']
    for col, h in enumerate(disc_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    discounts = [
        ['CAT-A', 15],
        ['CAT-B', 10],
        ['CAT-C', 5],
    ]
    for r, row_data in enumerate(discounts, 2):
        for c, val in enumerate(row_data, 1):
            ws2.cell(row=r, column=c, value=val)

    ws2.column_dimensions['A'].width = 16
    ws2.column_dimensions['B'].width = 14

    wb.save(OUTPUT)
    print(f'Initial file created: {OUTPUT}')

    # GUI-ready startup
    launch_gui(f'libreoffice --calc "{OUTPUT}"', delay_sec=2.0)
    print('GUI_READY: launched LibreOffice Calc with DISPLAY=:0')


create_initial()
