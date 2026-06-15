"""
Reward Script: Nested VLOOKUP formula verification
Task ID: calc_lf_017
Domain: libreoffice_calc
Scoring:
  Component 1 (0.3): C2 contains a formula (starts with '=')
  Component 2 (0.4): C2 formula contains nested VLOOKUP (VLOOKUP inside VLOOKUP)
  Component 3 (0.3): C2 formula references the Discounts sheet for the outer lookup
"""

import os
import re
import openpyxl

WORKDIR = '/home/user'
TASK_ID = 'calc_lf_017'


def verify_task(file_path):
    """
    Verify task completion with progressive scoring.
    Returns: float between 0.0 and 1.0
    """
    total_score = 0.0

    try:
        wb = openpyxl.load_workbook(file_path)
    except Exception as e:
        print(f"CRITICAL: Cannot load file {file_path}: {e}")
        print("REWARD: 0.0")
        return 0.0

    # Precondition: Products sheet must exist
    if 'Products' not in wb.sheetnames:
        print("FAIL: 'Products' sheet not found")
        print("REWARD: 0.0")
        return 0.0

    ws = wb['Products']
    c2_val = ws['C2'].value

    # Component 1: C2 contains a formula (0.3 points)
    # Initial env has C2 empty, golden has a formula
    try:
        if c2_val is not None and isinstance(c2_val, str) and c2_val.startswith('='):
            print(f"PASS: Component 1 — C2 contains a formula: {c2_val} (0.3 pts)")
            total_score += 0.3
        else:
            print(f"FAIL: Component 1 — C2 does not contain a formula, found: {repr(c2_val)}")
    except Exception as e:
        print(f"ERROR: Component 1 — {e}")

    # Component 2: C2 formula contains nested VLOOKUP (0.4 points)
    # The formula must have VLOOKUP as an argument inside another VLOOKUP
    try:
        if c2_val is not None and isinstance(c2_val, str):
            formula_upper = c2_val.upper().replace(' ', '')
            # Count VLOOKUP occurrences - need at least 2 for nesting
            vlookup_count = formula_upper.count('VLOOKUP(')
            if vlookup_count >= 2:
                # Verify actual nesting: inner VLOOKUP should be first arg of outer VLOOKUP
                # Pattern: =VLOOKUP(VLOOKUP(...),...)
                nested_pattern = r'VLOOKUP\s*\(\s*VLOOKUP\s*\('
                if re.search(nested_pattern, c2_val, re.IGNORECASE):
                    print(f"PASS: Component 2 — Nested VLOOKUP found ({vlookup_count} VLOOKUPs) (0.4 pts)")
                    total_score += 0.4
                else:
                    print(f"FAIL: Component 2 — Found {vlookup_count} VLOOKUPs but not properly nested")
            else:
                print(f"FAIL: Component 2 — Expected nested VLOOKUP (>=2), found {vlookup_count} VLOOKUP(s)")
        else:
            print(f"FAIL: Component 2 — C2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 2 — {e}")

    # Component 3: Outer VLOOKUP references the Discounts sheet (0.3 points)
    # The formula should reference Discounts sheet (e.g., Discounts.A2:B4 or Discounts!A2:B4)
    try:
        if c2_val is not None and isinstance(c2_val, str):
            formula_check = c2_val.upper().replace(' ', '')
            # LibreOffice uses dot notation (Discounts.A2:B4), Excel uses ! (Discounts!A2:B4)
            # Accept either format
            if 'DISCOUNTS.' in formula_check or 'DISCOUNTS!' in formula_check:
                print(f"PASS: Component 3 — Formula references Discounts sheet (0.3 pts)")
                total_score += 0.3
            else:
                print(f"FAIL: Component 3 — Formula does not reference Discounts sheet. Formula: {c2_val}")
        else:
            print(f"FAIL: Component 3 — C2 is not a formula string")
    except Exception as e:
        print(f"ERROR: Component 3 — {e}")

    final_score = min(total_score, 1.0)
    print(f"\nScore: {total_score}/1.0")
    print(f"REWARD: {final_score}")
    return final_score


# Entry point
file_path = f'{WORKDIR}/{TASK_ID}.xlsx'
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    print("REWARD: 0.0")
else:
    verify_task(file_path)
